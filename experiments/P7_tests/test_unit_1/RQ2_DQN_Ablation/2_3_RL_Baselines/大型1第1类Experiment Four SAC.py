import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
from torch.distributions import Normal
import numpy as np
import random
import time
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# === 配置（新范围：温度 1~300，电压 1~200，流量 1~5） ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'ACTION_DIM': 3,
    'MIN_VALUES': np.array([1, 1, 1], dtype=np.float32),
    'MAX_VALUES': np.array([300, 200, 5], dtype=np.float32),
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 3,
    'REPLAY_BATCH_SIZE': 64,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'ACTOR_LR': 3e-4,
    'CRITIC_LR': 3e-4,
    'ALPHA_LR': 3e-4,
    'GAMMA': 0.99,
    'TAU': 0.005,
    'NUM_RUNS': 20,
    'TOP_K_SAMPLES': 20,
    'TARGET_PATHS': [
        {7, 8, 12, 13, 14, 15, 20, 21, 22, 23, 25, 26, 27, 31, 32, 34, 35, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 50, 51,
         52, 56, 57, 58, 59, 61, 63, 64, 65, 66, 67, 68, 69, 71, 75, 76, 77, 78, 79, 80, 86, 90, 91, 93, 94, 95, 103, 104,
         105, 106, 107, 108, 110, 111, 112, 113, 114, 122, 123, 126, 127, 131, 132, 133, 143, 144, 148, 149, 150, 159, 160},
        {7, 8, 12, 13, 14, 15, 17, 18, 19, 21, 22, 23, 25, 26, 27, 31, 32, 34, 35, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46,
         50, 51, 52, 56, 57, 58, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 75, 76, 77, 78, 79, 80, 86, 90, 91, 93, 94, 95,
         103, 104, 105, 106, 107, 108, 110, 111, 112, 113, 114, 122, 123, 126, 127, 131, 132, 133, 143, 144, 159, 160},
        {7, 13, 14, 15, 20, 21, 22, 23, 26, 27, 31, 32, 34, 35, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51,
         52, 56, 57, 58, 59, 63, 64, 65, 66, 67, 68, 69, 70, 72, 75, 76, 77, 78, 79, 80, 86, 90, 91, 93, 94, 95, 103, 104,
         105, 109, 112, 113, 114, 122, 123, 126, 127, 131, 132, 133, 134, 135, 136, 143, 144, 148, 149, 150, 157, 158},
        {4, 7, 8, 9, 12, 13, 14, 15, 21, 22, 24, 31, 33, 38, 39, 40, 41, 42, 43, 44, 45, 46, 57, 58, 59, 63, 64, 65, 66, 67,
         68, 69, 71, 75, 76, 77, 78, 79, 80, 84, 85, 86, 87, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 106, 107,
         108, 109, 112, 113, 114, 118, 122, 123, 126, 127, 131, 132, 133, 148, 149, 150, 151, 152, 153, 159, 160},
        {7, 9, 13, 14, 15, 20, 21, 22, 29, 30, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 64, 65, 66,
         67, 68, 69, 75, 76, 77, 78, 79, 80, 84, 85, 86, 87, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 109, 112,
         113, 114, 118, 122, 123, 126, 127, 128, 131, 132, 133, 134, 136, 141, 151, 152, 153, 161, 162, 163},
        {7, 9, 13, 14, 15, 20, 21, 22, 24, 29, 30, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 47, 48, 49, 57, 58, 59, 64,
         65, 66, 67, 68, 69, 75, 76, 77, 84, 85, 86, 87, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 109, 112, 113,
         114, 118, 122, 123, 126, 127, 129, 131, 132, 133, 134, 135, 136, 151, 152, 153, 161, 162, 163},
        {7, 9, 13, 14, 15, 20, 21, 22, 24, 29, 30, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 64, 65,
         66, 67, 68, 69, 75, 76, 77, 78, 79, 80, 84, 85, 86, 87, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 109,
         112, 113, 114, 118, 122, 123, 126, 127, 130, 131, 132, 133, 151, 152, 153, 161, 162, 163},
        {7, 9, 13, 14, 15, 20, 21, 22, 29, 30, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 67, 68, 69,
         75, 76, 77, 78, 79, 80, 84, 85, 86, 87, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 109, 118, 122, 123,
         126, 127, 129, 131, 132, 133, 134, 135, 136, 140, 141, 142, 151, 152, 153, 161, 162, 163},
        {7, 13, 14, 15, 20, 23, 25, 26, 27, 31, 32, 34, 35, 37, 38, 39, 40, 44, 45, 46, 50, 51, 52, 56, 57, 58, 59, 61, 63,
         64, 65, 66, 67, 68, 69, 70, 72, 75, 76, 77, 78, 79, 80, 81, 82, 83, 86, 90, 91, 101, 102, 109, 112, 113, 114, 115,
         116, 117, 122, 123, 131, 132, 133, 134, 135, 136, 143, 144, 148, 149, 150, 157, 158},
        {7, 11, 13, 14, 15, 17, 18, 19, 23, 25, 26, 27, 31, 35, 37, 38, 39, 40, 44, 45, 46, 50, 51, 52, 56, 57, 58, 59, 60,
         61, 62, 63, 64, 65, 66, 70, 72, 78, 79, 80, 81, 82, 83, 86, 101, 102, 103, 104, 105, 110, 111, 115, 116, 117, 122,
         123, 131, 132, 133, 134, 135, 136, 143, 144, 146, 147, 148, 149, 150, 156, 157, 158},
        {5, 9, 13, 14, 15, 20, 28, 29, 30, 32, 34, 44, 45, 46, 53, 54, 60, 61, 62, 64, 65, 66, 70, 72, 73, 78, 79, 80, 81,
         82, 83, 84, 85, 86, 87, 88, 89, 96, 97, 98, 103, 104, 105, 110, 111, 115, 116, 117, 118, 122, 123, 128, 131, 132,
         133, 134, 136, 140, 141, 144, 148, 149, 150, 151, 152, 153, 157, 158, 161, 162, 163},
        {5, 9, 13, 14, 15, 20, 28, 29, 30, 32, 34, 44, 45, 46, 53, 54, 60, 61, 62, 64, 65, 66, 70, 72, 73, 74, 78, 79, 80,
         81, 82, 83, 84, 85, 86, 87, 88, 89, 96, 97, 98, 103, 104, 105, 110, 111, 115, 116, 117, 118, 122, 123, 128, 131,
         132, 133, 134, 136, 140, 141, 148, 149, 150, 151, 152, 153, 157, 158, 161, 162, 163},
        {4, 7, 8, 9, 12, 14, 21, 22, 31, 33, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 53, 54, 57, 58, 59, 63, 64, 65,
         66, 67, 68, 69, 71, 75, 76, 77, 78, 79, 80, 84, 85, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 106, 107,
         108, 109, 112, 113, 114, 118, 122, 123, 124, 126, 127, 131, 132, 133, 159, 160},
        {6, 13, 14, 15, 17, 18, 19, 23, 25, 26, 27, 31, 35, 37, 38, 39, 40, 44, 45, 46, 50, 51, 52, 56, 57, 58, 59, 60, 61,
         62, 63, 64, 65, 66, 70, 72, 78, 79, 80, 81, 82, 83, 86, 101, 102, 103, 104, 105, 110, 111, 115, 116, 117, 120, 122,
         123, 131, 132, 133, 143, 144, 146, 147, 148, 149, 150, 154, 155, 157, 158},
        {6, 13, 14, 15, 17, 18, 19, 23, 25, 26, 27, 31, 35, 37, 38, 39, 40, 44, 45, 46, 50, 51, 52, 56, 60, 61, 62, 63, 64,
         65, 66, 70, 72, 78, 79, 80, 81, 82, 83, 86, 101, 102, 103, 104, 105, 108, 110, 111, 115, 116, 117, 120, 122, 123,
         131, 132, 133, 141, 143, 144, 145, 148, 149, 150, 154, 155, 157, 158},
        {4, 7, 9, 10, 12, 21, 22, 31, 33, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 53, 54, 57, 58, 59, 63, 64, 65,
         66, 67, 68, 69, 71, 75, 76, 77, 78, 79, 80, 85, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 109, 112, 113,
         114, 118, 122, 123, 126, 127, 131, 132, 133, 148, 149, 150, 157, 158},
        {7, 13, 14, 15, 20, 23, 26, 27, 31, 32, 34, 35, 37, 38, 39, 40, 41, 42, 43, 50, 51, 52, 57, 58, 59, 65, 66, 67, 68,
         69, 70, 72, 75, 76, 77, 78, 79, 80, 81, 82, 83, 86, 90, 91, 101, 102, 109, 112, 113, 114, 115, 116, 117, 122, 123,
         125, 131, 132, 133, 134, 143, 144, 148, 149, 150, 157, 158},
        {7, 9, 13, 14, 15, 28, 31, 33, 38, 39, 40, 44, 45, 46, 53, 54, 55, 57, 58, 59, 63, 64, 65, 66, 67, 68, 69, 70, 72,
         78, 79, 80, 81, 82, 83, 84, 85, 86, 87, 88, 89, 103, 104, 105, 109, 115, 116, 117, 118, 122, 123, 131, 132, 133,
         144, 146, 147, 148, 149, 150, 151, 152, 153, 157, 158},
        {1, 2, 3, 7, 9, 13, 14, 15, 20, 21, 22, 29, 30, 33, 35, 36, 37, 41, 42, 43, 44, 45, 46, 53, 54, 56, 61, 78, 79, 80,
         84, 85, 86, 87, 88, 89, 103, 104, 105, 109, 118, 122, 123, 126, 127, 129, 131, 132, 133, 134, 135, 136, 140, 141,
         142, 146, 147, 151, 152, 153, 156, 162, 163},
        {1, 2, 3, 4, 6, 9, 13, 14, 15, 16, 20, 21, 22, 29, 30, 33, 53, 54, 61, 78, 79, 80, 84, 85, 86, 87, 88, 89, 109, 118,
         119, 120, 121, 122, 123, 126, 127, 129, 131, 132, 133, 134, 135, 136, 137, 138, 139, 140, 141, 142, 145, 151, 152,
         153, 154, 155, 162, 163},
        {5, 9, 13, 14, 15, 17, 18, 19, 21, 22, 32, 34, 35, 36, 37, 44, 45, 46, 53, 54, 60, 61, 62, 64, 65, 66, 73, 74, 84,
         85, 86, 87, 88, 89, 92, 96, 97, 98, 110, 111, 118, 122, 123, 126, 127, 129, 151, 152, 153},
    ]
}

# === 触发函数（流量压力密度混合规则集，与目标路径编号匹配） ===
def section1_flow_pressure_density_hybrid(x, y, z):
    b = {}
    triggered = set()

    # 规则1-10
    if (x > 190 and x < 210 and y > 90 and y < 110) != (x > 190 and x < 210 and y * 2 > 90 and y < 110): b[0] = 1
    if (x > 190 and x < 210 and y > 90 and y < 110) != (x > 190 and x < 210 and y * 4 > 90 and y < 110): b[1] = 2
    if (x > 190 and x < 210 and y > 90 and y < 110) != (x > 190 and x < 210 and y * 8 > 90 and y < 110): b[2] = 3
    if (y > 85 and y < 115 and z > 1.25 and z < 1.45) != (y > 85 and y < 115 and z > 1.25 and z < 15): b[3] = 4
    if (x > 195 and x < 205) != (x * 8 > 195 and x < 205): b[4] = 5
    if (x > 195 and x < 205) != (x > 195 and x * 8 < 205): b[5] = 6
    if (x > 195 and x < 205) != (x > 195 and x < 605): b[6] = 7
    if (y > 95 and y < 105) != (y > 95 and y * 8 < 105): b[7] = 8
    if (z > 1.3 and z < 1.4) != (z > 1.3 and z < 14): b[8] = 9
    if (z > 1 and z < 3.4) != (z > 1 and z < 4.4): b[9] = 10

    # 规则11-20
    if (abs(x - 200) < 8) != (abs(x - 200) < 18): b[10] = 11
    if (abs(y - 100) < 8) != (abs(y ** 2 - 100) < 8): b[11] = 12
    if (abs(z - 1.35) < 0.08) != (abs(z - 1.35) < 1.08): b[12] = 13
    if (abs(z - 1.35) < 0.08) != (abs(z - 1.35) < 2.08): b[13] = 14
    if (abs(z - 1.35) < 1.08) != (abs(z + 1.35) < 1.08): b[14] = 15
    if (x > 188 and x < 212 and y > 88 and y < 112 and z > 1.28 and z < 1.42) != (
            x > 188 and x < 212 and y > 88 and y < 112 and z > 1.28 and z < 9.42): b[15] = 16
    if ((x + y + z * 100) > 420 and (x + y + z * 100) < 460) != ((x + y + z * 100) > 720 and (x + y + z * 100) < 460):
        b[16] = 17
    if ((x + y + z * 100) > 420 and (x + y + z * 100) < 460) != ((x + y + z * 100) > 420 and (x + y + z * 100) < 260):
        b[17] = 18
    if ((x + y + z * 100) > 420 and (x + y + z * 100) < 460) != ((x + y + z * 100) > 420 and (x + y + z * 100) < 360):
        b[18] = 19
    if ((x + y + z * 100) > 420 and (x + y + z * 100) < 460) != ((x + y + z * 100) > 420 and (x + y + z * 100) < 560):
        b[19] = 20

    # 规则21-30
    if (x / y > 1.8 and x / y < 2.2) != (x / y > 1.8 and x / y < 22): b[20] = 21
    if (x / y > 1.8 and x / y < 2.2) != (x / y > 1.8 and x / y < 12.2): b[21] = 22
    if (x / z > 130 and x / z < 170) != (x / z > 130 and x / z < 1170): b[22] = 23
    if (x / z > 130 and x / z < 170) != (x / z > 130 and x / z < 120): b[23] = 24
    if (x / z > 130 and x / z < 170) != (x / z > 130 and x / z < 270): b[24] = 25
    if (y / z > 60 and y / z < 90) != (y / z > 60 and y / z < 910): b[25] = 26
    if (y / z > 60 and y / z < 90) != (y / z > 60 and y / z < 190): b[26] = 27
    if (y / z > 60 and y / z < 90) != (y / z > 60 and y / z < 60): b[27] = 28
    if (x * y > 16000 and x * y < 24000) != (x * y > 16000 and x * y < 2400): b[28] = 29
    if (x * y > 16000 and x * y < 24000) != (x * y > 16000 and x * y < 12400): b[29] = 30

    # 规则31-40
    if (x * y > 16000 and x * y < 24000) != (x * y > 16000 and 22 * y < 24000): b[30] = 31
    if (x * z > 240 and x * z < 320) != (x * z > 240 and x * 22 < 320): b[31] = 32
    if (x * z > 240 and x * z < 320) != (x * z > 240 and 22 * z < 320): b[32] = 33
    if (x * z > 240 and x * z < 320) != (x * z > 240 and x * 99 < 320): b[33] = 34
    if (y * z > 100 and y * z < 170) != (y * z > 100 and y * 22 < 170): b[34] = 35
    if (y * z > 100 and y * z < 170) != (y * z > 100 and 122 * z < 170): b[35] = 36
    if (y * z > 100 and y * z < 170) != (y * z > 100 and y * 88 < 170): b[36] = 37
    if ((x + y) / 2 > 145 and (x + y) / 2 < 155) != ((x + y) / 2 > 145 and (x + y) / 4 < 155): b[37] = 38
    if ((x + y) / 2 > 145 and (x + y) / 2 < 155) != ((x + y) / 2 > 145 and (x + y) / 2 < 455): b[38] = 39
    if ((x + y) / 2 > 145 and (x + y) / 2 < 155) != ((x + y) / 2 > 145 and (x + y) / 2 < 855): b[39] = 40

    # 规则41-50
    if (x - y > 60 and x - y < 120) != (x - y > 60 and x - y < 420): b[40] = 41
    if (x - y > 60 and x - y < 120) != (x - y > 60 and x - y < 620): b[41] = 42
    if (x - y > 60 and x - y < 120) != (x - y > 60 and x - y < 820): b[42] = 43
    if (abs(x / y - 2.0) < 0.3) != (abs(x / y - 2.0) < 1.3): b[43] = 44
    if (abs(x / y - 2.0) < 0.3) != (abs(x / y - 2.0) < 2.3): b[44] = 45
    if (abs(x / y - 2.0) < 0.3) != (abs(x / y - 2.0) < 55.3): b[45] = 46
    if (x / (y + 50) > 1.2 and x / (y + 50) < 1.8) != (x / (y + 50) > 1.2 and x / (y + 50) < 18): b[46] = 47
    if (x / (y + 50) > 1.2 and x / (y + 50) < 1.8) != (x / (y + 50) > 1.2 and x / (y + 50) < 221.8): b[47] = 48
    if (x / (y + 50) > 1.2 and x / (y + 50) < 1.8) != (x / (y + 50) > 1.2 and x / (y + 50) < 118): b[48] = 49
    if (y / (z * 50) > 1.2 and y / (z * 50) < 1.8) != (y / (z * 50) > 1.2 and y / (z * 50) < 1118): b[49] = 50

    # 规则51-60
    if (y / (z * 50) > 1.2 and y / (z * 50) < 1.8) != (y / (z * 50) > 1.2 and y / (z * 50) < 221.8): b[50] = 51
    if (y / (z * 50) > 1.2 and y / (z * 50) < 1.8) != (y / (z * 50) > 1.2 and y / (z * 50) < 122.8): b[51] = 52
    if (z / (x / 200) > 1.2 and z / (x / 200) < 1.8) != (z / (x / 200) > 1.2 and z / (x / 200) < 111.8): b[52] = 53
    if (z / (x / 200) > 1.2 and z / (x / 200) < 1.8) != (z / (x / 200) > 1.2 and z / (x / 200) < 18): b[53] = 54
    if (z / (x / 200) > 1.2 and z / (x / 200) < 1.8) != (z / (x / 200) > 1.2 and z / (x / 200) < 1.811): b[54] = 55
    if ((x * y * z) > 25000 and (x * y * z) < 35000) != ((x * y * z) > 25000 and (x * y * z) < 300): b[55] = 56
    if ((x + y + z) / 3 > 108 and (x + y + z) / 3 < 118) != ((x + y + z) / 3 > 108 and (x + y + z) / 5 < 118): b[56] = 57
    if ((x + y + z) / 3 > 108 and (x + y + z) / 3 < 118) != ((x + y + z) / 3 > 108 and (x + y + z) / 9 < 118): b[57] = 58
    if ((x + y + z) / 3 > 108 and (x + y + z) / 3 < 118) != ((x + y + z) / 3 > 108 and (x + y + z) / 8 < 118): b[58] = 59
    if (x ** 2 + y ** 2 + z ** 2 * 10000 > 50000) != (x ** 2 + y ** 2 + z ** 2 * 10000 > 80000): b[59] = 60

    # 规则61-70
    if (x ** 2 + y ** 2 + z ** 2 * 10000 > 50000) != (x ** 2 + y ** 2 + z ** 2 * 10000 > 99000): b[60] = 61
    if (x ** 2 + y ** 2 + z ** 2 * 10000 > 50000) != (x ** 2 + y ** 2 + z ** 2 * 10000 > 88000): b[61] = 62
    if ((x - 180) * (y - 80) > 200 and (x - 180) * (y - 80) < 1000) != (
            (x - 180) * (y - 80) > 200 and (x - 180) * (y - 80) < 4000): b[62] = 63
    if (abs((x + y) - 300) < 20) != (abs((x + y) - 300) < 120): b[63] = 64
    if (abs((x + y) - 300) < 20) != (abs((x + y) - 300) < 210): b[64] = 65
    if (abs((x + y) - 300) < 20) != (abs((x + y) - 300) < 2990): b[65] = 66
    if (x / 200 > 0.9 and x / 200 < 1.1) != (x / 200 > 0.9 and x / 200 < 11): b[66] = 67
    if (x / 200 > 0.9 and x / 200 < 1.1) != (x / 200 > 0.9 and x / 200 < 8.1): b[67] = 68
    if (x / 200 > 0.9 and x / 200 < 1.1) != (x / 200 > 0.9 and x / 200 < 99.1): b[68] = 69
    if (y / 100 > 0.9 and y / 100 < 1.1) != (y / 100 > 0.9 and y / 100 < 11): b[69] = 70

    # 规则71-80
    if (y / 100 > 0.9 and y / 100 < 1.1) != (y / 100 > 0.9 and y / 10 < 1.1): b[70] = 71
    if (y / 100 > 0.9 and y / 100 < 1.1) != (y / 100 > 0.9 and y / 200 < 1.1): b[71] = 72
    if (x < 175) != (x < 115): b[72] = 73
    if (x < 175) != (x < 125): b[73] = 74
    if (x > 225) != (x > 2215): b[74] = 75
    if (x > 225) != (x > 2299): b[75] = 76
    if (x > 225) != (x > 1225): b[76] = 77
    if (y < 75) != (y < 175): b[77] = 78
    if (y < 75) != (y < 715): b[78] = 79
    if (y < 75) != (y < 751): b[79] = 80

    # 规则81-90
    if (y > 125) != (y > 925): b[80] = 81
    if (y > 125) != (y > 1115): b[81] = 82
    if (y > 125) != (y > 1215): b[82] = 83
    if (z < 1.15) != (z < 3.85): b[83] = 84
    if (z < 1.15) != (z < 4.5): b[84] = 85
    if (z < 2.15) != (z < 0.15): b[85] = 86
    if (z > 2.55) != (z > 1.25): b[86] = 87
    if (z > 1.55) != (z > 7.45): b[87] = 88
    if (z > 1.55) != (z > 165): b[88] = 89
    if (x < 170 or x > 230) != (x < 170 or x > 1230): b[89] = 90

    # 规则91-100
    if (x < 170 or x > 230) != (x < 170 or x > 2130): b[90] = 91
    if (x < 165 and y < 75) != (x < 165 and y * 8 < 75): b[91] = 92
    if (x > 235 and y > 125) != (x > 235 and y ** 9 > 125): b[92] = 93
    if (x > 235 and y > 125) != (x > 235 and y ** 2 > 125): b[93] = 94
    if (x > 235 and y > 125) != (x > 235 and y ** 4 > 125): b[94] = 95
    if (x < 170 and z < 1.15) != (x < 170 and z < 4.15): b[95] = 96
    if (x < 170 and z < 1.15) != (x < 170 and z < 9.15): b[96] = 97
    if (x < 170 and z < 1.15) != (x < 170 and z < 94.15): b[97] = 98
    if (x > 230 and z > 1.55) != (x > 230 and z > 155): b[98] = 99
    if (x > 230 and z > 1.55) != (x > 230 and z > 4.55): b[99] = 100

    # 规则101-110
    if (y > 125 and z > 1.55) != (y > 125 and z * 2 > 1.55): b[100] = 101
    if (y > 125 and z > 1.55) != (y > 125 and z * 6 > 1.55): b[101] = 102
    if (abs(x / y - 2.0) > 0.5) != (abs(x / y - 2.0) > 10.5): b[102] = 103
    if (abs(x / y - 2.0) > 0.5) != (abs(x / y - 2.0) > 120.5): b[103] = 104
    if (abs(x / y - 2.0) > 0.5) != (abs(x / y - 2.0) > 5): b[104] = 105
    if (x * y < 14000 or x * y > 26000) != (x * y < 14000 or x * y > 28000): b[105] = 106
    if (x * y < 14000 or x * y > 26000) != (x * y < 14000 or x * y > 27000): b[106] = 107
    if (x * y < 14000 or x * y > 26000) != (x * y < 14000 or x * y > 30000): b[107] = 108
    if ((x + y + z * 100) < 400 or (x + y + z * 100) > 480) != ((x + y + z * 100) < 400 or (x + y + z * 100) > 4180): b[108] = 109
    if ((x + y + z * 100) < 400 or (x + y + z * 100) > 480) != ((x + y + z * 100) < 400 or (x + y + z * 1100) > 480): b[109] = 110

    # 规则111-120
    if ((x + y + z * 100) < 400 or (x + y + z * 100) > 480) != ((x + y + z * 100) < 400 or (x + y + z * 800) > 480): b[110] = 111
    if (x < 160 or x > 240) != (x < 160 or x > 1240): b[111] = 112
    if (x < 160 or x > 240) != (x < 160 or x > 2140): b[112] = 113
    if (x < 160 or x > 240) != (x < 160 or x > 2410): b[113] = 114
    if (y < 60 or y > 140) != (y < 60 or y > 1410): b[114] = 115
    if (y < 60 or y > 140) != (y < 60 or y > 1420): b[115] = 116
    if (y < 60 or y > 140) != (y < 60 or y > 1770): b[116] = 117
    if (z > 1.33 and z < 1.37) != (z > 1.33 and z < 137): b[117] = 118
    if (abs(x - 200) < 3 and abs(y - 100) < 3) != (abs(x - 200) < 3 and abs(y - 100) < 13): b[118] = 119
    if (abs(x - 200) < 3 and abs(y - 100) < 3) != (abs(x - 200) < 3 and abs(y - 100) < 322): b[119] = 120

    # 规则121-130
    if (abs(x - 200) < 3 and abs(y - 100) < 3) != (abs(x - 200) < 3 and abs(y - 100) < 31): b[120] = 121
    if (abs(z - 1.35) < 0.03) != (abs(z - 1.35) < 4.03): b[121] = 122
    if (abs(z - 1.35) < 0.03) != (abs(z - 1.35) < 8.03): b[122] = 123
    if (abs(z - 1.35) < 2.03) != (abs(z - 1.35) < 1.3): b[123] = 124
    if (x / y > 1.95 and x / y < 2.05) != (x / y > 125 and x / y < 2.05): b[124] = 125
    if (x / y > 1.95 and x / y < 2.05) != (x / y > 1.95 and x / y < 205): b[125] = 126
    if (x / y > 1.95 and x / y < 2.05) != (x / y > 1.95 and x / y < 2105): b[126] = 127
    if (x * y > 19000 and x * y < 21000) != (x * y > 19000 and x * y < 1000): b[127] = 128
    if (x * y > 19000 and x * y < 21000) != (x * y > 1900 and x * y < 21000): b[128] = 129
    if (x * y > 19000 and x * y < 21000) != (x * y > 19000 and x * y < 21090): b[129] = 130

    # 规则131-140
    if ((x + y + z * 100) > 430 and (x + y + z * 100) < 440) != ((x + y + z * 100) > 430 and (x + y + z * 100) < 4140): b[130] = 131
    if ((x + y + z * 100) > 430 and (x + y + z * 100) < 440) != ((x + y + z * 100) > 430 and (x + y + z * 100) < 1440): b[131] = 132
    if ((x + y + z * 100) > 430 and (x + y + z * 100) < 440) != ((x + y + z * 100) > 430 and (x + y + z * 100) < 2440): b[132] = 133
    if ((x * y * z) > 29000 and (x * y * z) < 31000) != ((x * y * z) > 29000 and (x * y * z) < 41000): b[133] = 134
    if ((x * y * z) > 29000 and (x * y * z) < 31000) != ((x * y * z) > 29000 and (x * y * z) < 38000): b[134] = 135
    if ((x * y * z) > 29000 and (x * y * z) < 31000) != ((x * y * z) > 29000 and (x * y * z) < 39000): b[135] = 136
    if (((x - 200) ** 2 + (y - 100) ** 2) < 25) != (((x - 200) ** 2 + (y - 100) ** 2) < 125): b[136] = 137
    if (((x - 200) ** 2 + (y - 100) ** 2) < 25) != (((x - 200) ** 2 + (y - 100) ** 2) < 215): b[137] = 138
    if (((x - 200) ** 2 + (y - 100) ** 2) < 25) != (((x - 200) ** 2 + (y - 100) ** 2) < 251): b[138] = 139
    if (abs((x + y) / 2 - 150) < 2) != (abs((x + y) / 2 - 150) < 12): b[139] = 140

    # 规则141-150
    if (abs((x + y) / 2 - 150) < 2) != (abs((x + y) / 2 - 150) < 21): b[140] = 141
    if (abs((x + y) / 2 - 150) < 2) != (abs((x + y) / 2 - 150) < 9): b[141] = 142
    if (x / (z * 100) > 1.45 and x / (z * 100) < 1.55) != (x / (z * 100) > 1.45 and x / (z * 100) < 11.55): b[142] = 143
    if (y / (z * 50) > 1.45 and y / (z * 50) < 1.55) != (y / (z * 50) > 1.45 and y / (z * 50) < 55): b[143] = 144
    if ((x - 180) / 40 > 0.45 and (x - 180) / 40 < 0.55) != ((x - 180) / 40 > 0.45 and (x - 180) / 4 < 0.55): b[144] = 145
    if ((x - 180) / 40 > 0.45 and (x - 180) / 40 < 0.55) != ((x - 180) / 40 > 0.45 and (x - 180) / 90 < 0.55): b[145] = 146
    if ((x - 180) / 40 > 0.45 and (x - 180) / 40 < 0.55) != ((x - 180) / 40 > 0.45 and (x - 180) / 80 < 0.55): b[146] = 147
    if ((y - 80) / 40 > 0.45 and (y - 80) / 40 < 0.55) != ((y - 80) / 40 > 0.45 and (y - 80) / 40 < 22.55): b[147] = 148
    if ((y - 80) / 40 > 0.45 and (y - 80) / 40 < 0.55) != ((y - 80) / 40 > 0.45 and (y - 80) / 40 < 2.55): b[148] = 149
    if ((y - 80) / 40 > 0.45 and (y - 80) / 40 < 0.55) != ((y - 80) / 40 > 0.45 and (y - 80) / 40 < 6.55): b[149] = 150

    # 规则151-160
    if ((z - 1.2) / 0.3 > 0.45 and (z - 1.2) / 0.3 < 0.55) != ((z - 1.2) / 0.3 > 0.45 and (z - 1.2) / 3 < 0.55): b[150] = 151
    if ((z - 1.2) / 0.3 > 0.45 and (z - 1.2) / 0.3 < 0.55) != ((z - 1.2) / 0.3 > 0.45 and (z - 1.2) / 0.3 < 5.55): b[151] = 152
    if ((z - 1.2) / 0.3 > 0.45 and (z - 1.2) / 0.3 < 0.55) != ((z - 1.2) / 0.3 > 0.45 and (z - 1.2) / 3 < 0.55): b[152] = 153
    if (min(x - 180, 220 - x) > 15) != (min(x - 180, 220 - x) > 152): b[153] = 154
    if (min(x - 180, 220 - x) > 15) != (min(x - 180, 220 - x) > 115): b[154] = 155
    if (min(x - 180, 220 - x) > 15) != (min(x - 180, 220 - x) > 5): b[155] = 156
    if (min(y - 80, 120 - y) > 15) != (min(y - 80, 120 + y) > 15): b[156] = 157
    if (min(y - 80, 120 - y) > 15) != (min(y - 80, 120 * y) > 15): b[157] = 158
    if (min(y - 80, 120 - y) > 15) != (min(y - 80, 120 - y) > 95): b[158] = 159
    if (min(y - 80, 120 - y) > 15) != (min(y - 80, 120 - y) > 35): b[159] = 160

    # 规则161-163
    if (abs(x * y - 20000) < 500) != (abs(x * y - 20000) < 1500): b[160] = 161
    if (abs(x * y - 20000) < 500) != (abs(x * y - 20000) < 5100): b[161] = 162
    if (abs(x * y - 20000) < 500) != (abs(x * y - 20000) < 5010): b[162] = 163

    for key, value in b.items():
        triggered.add(value)
    return triggered

# === 绑定执行函数 ===
execute_Tr = section1_flow_pressure_density_hybrid

# === 状态处理辅助函数 ===
def clip_state(state):
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return np.clip(state, min_vals, max_vals)

def normalize_state(state):
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return 2 * (state - min_vals) / (max_vals - min_vals) - 1

def denormalize_state(normalized_state):
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return (normalized_state + 1) * (max_vals - min_vals) / 2 + min_vals

def coverage_similarity(triggered, target_path):
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0
    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)

def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']
    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']
    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']
    return reward

# === SAC Actor ===
class GaussianPolicy(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dim=256):
        super(GaussianPolicy, self).__init__()
        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.mean = nn.Linear(hidden_dim, action_dim)
        self.log_std = nn.Linear(hidden_dim, action_dim)
        self.action_scale = 8.0
        self.action_bias = 0.0

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))
        mean = self.mean(x)
        log_std = self.log_std(x)
        log_std = torch.clamp(log_std, min=-20, max=2)
        return mean, log_std

    def sample(self, state):
        mean, log_std = self.forward(state)
        std = log_std.exp()
        normal = Normal(mean, std)
        x_t = normal.rsample()
        action = torch.tanh(x_t) * self.action_scale + self.action_bias
        log_prob = normal.log_prob(x_t)
        log_prob -= torch.log(self.action_scale * (1 - torch.tanh(x_t).pow(2)) + 1e-6)
        log_prob = log_prob.sum(1, keepdim=True)
        mean = torch.tanh(mean) * self.action_scale + self.action_bias
        return action, log_prob, mean

# === SAC Critic ===
class QNetwork(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dim=256):
        super(QNetwork, self).__init__()
        self.fc1 = nn.Linear(state_dim + action_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, 1)
        self.fc4 = nn.Linear(state_dim + action_dim, hidden_dim)
        self.fc5 = nn.Linear(hidden_dim, hidden_dim)
        self.fc6 = nn.Linear(hidden_dim, 1)

    def forward(self, state, action):
        xu = torch.cat([state, action], 1)
        x1 = F.relu(self.fc1(xu))
        x1 = F.relu(self.fc2(x1))
        q1 = self.fc3(x1)
        x2 = F.relu(self.fc4(xu))
        x2 = F.relu(self.fc5(x2))
        q2 = self.fc6(x2)
        return q1, q2

# === 经验回放缓冲区（包含路径和相似度信息） ===
class EnhancedReplayBuffer:
    def __init__(self, capacity=50000):
        self.buffer = deque(maxlen=capacity)
        self.experience_info = deque(maxlen=capacity)

    def push(self, state, action, reward, next_state, done, path_idx, similarity):
        self.buffer.append((state, action, reward, next_state, done))
        self.experience_info.append({'path_idx': path_idx, 'similarity': similarity})

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return None
        batch = random.sample(self.buffer, batch_size)
        state, action, reward, next_state, done = map(np.stack, zip(*batch))
        return (
            torch.FloatTensor(state).to(device),
            torch.FloatTensor(action).to(device),
            torch.FloatTensor(reward).unsqueeze(1).to(device),
            torch.FloatTensor(next_state).to(device),
            torch.FloatTensor(done).unsqueeze(1).to(device)
        )

    def get_top_k_per_path(self, num_paths, k=20):
        path_samples = {i: [] for i in range(num_paths)}
        for idx, info in enumerate(self.experience_info):
            path_idx = info['path_idx']
            similarity = info['similarity']
            path_samples[path_idx].append((idx, similarity, self.buffer[idx]))

        top_k_results = {}
        for path_idx in range(num_paths):
            samples = path_samples[path_idx]
            if len(samples) == 0:
                top_k_results[path_idx] = []
                continue
            samples.sort(key=lambda x: x[1], reverse=True)
            top_k = samples[:k]
            top_k_results[path_idx] = []
            for sample in top_k:
                normalized_state = sample[2][0]
                original_state = denormalize_state(normalized_state)
                original_state_rounded = np.round(original_state).astype(int)
                x, y, z = original_state_rounded
                triggered = execute_Tr(x, y, z)
                top_k_results[path_idx].append({
                    'state': original_state_rounded,
                    'similarity': sample[1],
                    'triggered': triggered
                })
        return top_k_results

    def __len__(self):
        return len(self.buffer)

# === SAC Agent ===
class SACAgent:
    def __init__(self, state_dim=3, action_dim=3):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.policy = GaussianPolicy(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.policy_optimizer = optim.Adam(self.policy.parameters(), lr=EXPERIMENT_CONFIG['ACTOR_LR'])
        self.critic = QNetwork(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_target = QNetwork(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_target.load_state_dict(self.critic.state_dict())
        self.critic_optimizer = optim.Adam(self.critic.parameters(), lr=EXPERIMENT_CONFIG['CRITIC_LR'])
        self.target_entropy = -action_dim
        self.log_alpha = torch.zeros(1, requires_grad=True, device=device)
        self.alpha_optimizer = optim.Adam([self.log_alpha], lr=EXPERIMENT_CONFIG['ALPHA_LR'])
        self.replay_buffer = EnhancedReplayBuffer()
        self.replay_train_count = 0

    def get_action(self, state, deterministic=False):
        normalized_state = normalize_state(state)
        state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)
        with torch.no_grad():
            if deterministic:
                _, _, action = self.policy.sample(state_tensor)
            else:
                action, _, _ = self.policy.sample(state_tensor)
        action = action.cpu().numpy()[0]
        return action

    def store_experience(self, state, action, reward, next_state, done, path_idx, similarity):
        normalized_state = normalize_state(state)
        normalized_next_state = normalize_state(next_state)
        self.replay_buffer.push(
            normalized_state, action, reward,
            normalized_next_state, done, path_idx, similarity
        )

    def replay_train(self):
        batch = self.replay_buffer.sample(EXPERIMENT_CONFIG['REPLAY_BATCH_SIZE'])
        if batch is None:
            return
        state, action, reward, next_state, done = batch

        with torch.no_grad():
            next_action, next_log_prob, _ = self.policy.sample(next_state)
            q1_next, q2_next = self.critic_target(next_state, next_action)
            q_next = torch.min(q1_next, q2_next)
            target_q = reward + (1 - done) * EXPERIMENT_CONFIG['GAMMA'] * (
                    q_next - self.log_alpha.exp() * next_log_prob)

        q1, q2 = self.critic(state, action)
        critic_loss = F.mse_loss(q1, target_q) + F.mse_loss(q2, target_q)

        self.critic_optimizer.zero_grad()
        critic_loss.backward()
        self.critic_optimizer.step()

        new_action, log_prob, _ = self.policy.sample(state)
        q1_new, q2_new = self.critic(state, new_action)
        q_new = torch.min(q1_new, q2_new)
        policy_loss = (self.log_alpha.exp() * log_prob - q_new).mean()

        self.policy_optimizer.zero_grad()
        policy_loss.backward()
        self.policy_optimizer.step()

        alpha_loss = -(self.log_alpha * (log_prob + self.target_entropy).detach()).mean()
        self.alpha_optimizer.zero_grad()
        alpha_loss.backward()
        self.alpha_optimizer.step()

        for param, target_param in zip(self.critic.parameters(), self.critic_target.parameters()):
            target_param.data.copy_(
                EXPERIMENT_CONFIG['TAU'] * param.data + (1 - EXPERIMENT_CONFIG['TAU']) * target_param.data)

        self.replay_train_count += 1
        if self.replay_train_count % 2 == 0:
            alpha_value = self.log_alpha.exp().item()
            print(f"  -> SAC 更新 (第 {self.replay_train_count} 次), Alpha={alpha_value:.4f}")

# === 性能计算函数 ===
def calculate_run_performance(run_idx, sac_results, training_time, total_steps, update_count, agent):
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)
    all_similarities = []
    total_samples = 0
    all_rewards = []
    total_reward = 0

    for path_idx in range(num_paths):
        samples = sac_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']
            total_reward += reward
            all_rewards.append(reward)
            all_similarities.append(similarity)
            total_samples += 1

    average_reward = total_reward / total_samples if total_samples > 0 else 0
    convergence = np.mean(all_similarities) if all_similarities else 0
    environment_adaptability = 1 / (np.std(all_similarities) + 1e-8) if len(all_similarities) > 1 else 0
    generalization_ability = convergence
    computational_efficiency = total_steps / training_time if training_time > 0 else 0
    policy_update_frequency = update_count / training_time if training_time > 0 else 0
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0

    return {
        'Run': run_idx + 1,
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }

# === Excel 导出函数 ===
def export_to_excel(all_sac_results, all_performance_data, target_paths, output_path="SAC_20_run.xlsx"):
    print("\n正在导出数据到 Excel...")
    all_sac_summary_data = []
    all_sac_detailed_data = []

    for run_idx, (sac_results, performance_data) in enumerate(zip(all_sac_results, all_performance_data)):
        sac_summary_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = sac_results[path_idx]
            if len(samples) == 0:
                sac_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue
            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = 'Yes' if perfect_count > 0 else 'No'
            sac_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })
        all_sac_summary_data.extend(sac_summary_data)

        sac_detailed_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = sac_results[path_idx]
            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']
                sac_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Sample ID': sample_idx + 1,
                    'X(温度)': int(state[0]),
                    'Y(电压)': int(state[1]),
                    'Z(流量)': int(state[2]),
                    'Similarity': round(similarity, 4),
                    'Perfect Coverage': 'Yes' if abs(similarity - 1.0) < 0.001 else 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path))),
                    'Triggered Rules': ', '.join(map(str, sorted(triggered))),
                    'Hit Rule Count': len(target_path.intersection(triggered)),
                    'Target Rule Count': len(target_path)
                })
        all_sac_detailed_data.extend(sac_detailed_data)

    sac_summary_df = pd.DataFrame(all_sac_summary_data)
    sac_detailed_df = pd.DataFrame(all_sac_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        sac_summary_df.to_excel(writer, sheet_name='SACPath', index=False)
        sac_detailed_df.to_excel(writer, sheet_name='SACDetailed Sample Data', index=False)
        selected_columns = [
            'Run', 'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df[selected_columns].to_excel(writer, sheet_name='Metric', index=False)

        workbook = writer.book
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        perfect_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        ws1 = writer.sheets['SACPath']
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row_idx in range(2, ws1.max_row + 1):
            if ws1.cell(row_idx, 9).value == 'Yes':
                for col_idx in range(1, ws1.max_column + 1):
                    ws1.cell(row_idx, col_idx).fill = perfect_fill
        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 18
        ws1.column_dimensions['D'].width = 15
        ws1.column_dimensions['E'].width = 18
        ws1.column_dimensions['F'].width = 15
        ws1.column_dimensions['G'].width = 15
        ws1.column_dimensions['H'].width = 15
        ws1.column_dimensions['I'].width = 18
        ws1.column_dimensions['J'].width = 50

        ws2 = writer.sheets['SACDetailed Sample Data']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 18
        ws2.column_dimensions['I'].width = 40
        ws2.column_dimensions['J'].width = 40
        ws2.column_dimensions['K'].width = 18
        ws2.column_dimensions['L'].width = 18

        ws3 = writer.sheets['Metric']
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws3.column_dimensions[col].width = 20

    print(f"文件已保存到: {output_path}")
    print(f"  - Sheet1: SACPath 统计 ({len(all_sac_summary_data)} 条)")
    print(f"  - Sheet2: SACDetailed Sample Data ({len(all_sac_detailed_data)} 条)")
    print(f"  - Sheet3: Metric ({len(all_performance_data)} 条)")

# === 单次训练流程 ===
def train_sac_workflow():
    print("=" * 80)
    print("SAC 训练 - 流量压力密度混合规则集")
    print("状态范围: X(温度) 1~300, Y(电压) 1~200, Z(流量) 1~5")
    print("相似度 = 命中规则数 / 目标路径规则数")
    print("=" * 80)

    agent = SACAgent()
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n生成初始样本: {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']} / 路径")
    path_samples = {}
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']

    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            state = np.random.uniform(min_vals, max_vals).astype(np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  路径 {path_idx + 1}/{num_paths}: {len(samples)} 个初始样本")

    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size

    print(f"\n批次大小: {batch_size}, 单样本迭代步数: {EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(f"总批次数: {num_batches} /路径 × {num_paths} 路径 = {num_batches * num_paths}")
    print("-" * 80)

    for batch_idx in range(num_batches):
        print(f"\n批次 {batch_idx + 1}/{num_batches}")

        for path_idx in range(num_paths):
            target_path = target_paths[path_idx]
            batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]

            batch_rewards = []
            batch_similarities = []

            for initial_state in batch_samples:
                state = initial_state.copy()
                episode_reward = 0
                final_similarity = 0

                for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                    action = agent.get_action(state, deterministic=False)
                    next_state = state + action
                    next_state = clip_state(next_state)

                    x, y, z = next_state
                    triggered = execute_Tr(x, y, z)
                    reward = unified_reward_function(triggered, target_path)
                    similarity = coverage_similarity(triggered, target_path)

                    done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)

                    agent.store_experience(
                        state, action, reward, next_state, done,
                        path_idx, similarity
                    )

                    state = next_state
                    episode_reward += reward
                    final_similarity = similarity
                    total_steps += 1

                batch_rewards.append(episode_reward)
                batch_similarities.append(final_similarity)

            avg_reward = np.mean(batch_rewards)
            avg_similarity = np.mean(batch_similarities)
            print(f"  路径 {path_idx + 1}: 平均奖励={avg_reward:.2f}, 平均相似度={avg_similarity:.4f}")

        print(f"\n  执行 SAC 更新...")
        agent.replay_train()
        print(f"  经验池大小: {len(agent.replay_buffer)}")

    training_time = time.time() - start_time

    print("\n" + "=" * 80)
    print(f"单次运行训练完成! 总耗时: {training_time:.2f} 秒, 总交互步数: {total_steps}")
    print(f"经验池大小: {len(agent.replay_buffer)}")
    print(f"SAC 更新次数: {agent.replay_train_count}")
    print("=" * 80)

    print(f"\n提取每条路径的 Top-{EXPERIMENT_CONFIG['TOP_K_SAMPLES']} 高相似度样本...")
    top_k_results = agent.replay_buffer.get_top_k_per_path(num_paths, EXPERIMENT_CONFIG['TOP_K_SAMPLES'])

    return agent, top_k_results, training_time, total_steps, agent.replay_train_count

# === 主程序 ===
def main():
    print("\n" + "=" * 80)
    print("SAC 多轮实验 - 流量压力密度混合规则测试")
    print("参数范围: 温度 1~300, 电压 1~200, 流量 1~5")
    print("=" * 80)

    all_sac_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'='*80}")
        print(f"开始第 {run_idx + 1}/{EXPERIMENT_CONFIG['NUM_RUNS']} 次独立运行")
        print(f"{'='*80}")

        sac_agent, sac_results, training_time, total_steps, update_count = train_sac_workflow()
        performance_data = calculate_run_performance(
            run_idx, sac_results, training_time, total_steps, update_count, sac_agent
        )

        all_sac_results.append(sac_results)
        all_performance_data.append(performance_data)

        print(f"\n第 {run_idx + 1} 次运行完成!")
        print(f"  总奖励: {performance_data['Total Reward']}")
        print(f"  平均奖励: {performance_data['Average Reward']}")
        print(f"  平均相似度: {performance_data['Average Similarity']}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"SAC_flow_pressure_density_20run_{timestamp}.xlsx"
    export_to_excel(all_sac_results, all_performance_data, target_paths, output_path)

    print("\n" + "=" * 80)
    print(f"{EXPERIMENT_CONFIG['NUM_RUNS']} 次运行整体统计摘要")
    print("=" * 80)

    total_rewards = [p['Total Reward'] for p in all_performance_data]
    average_rewards = [p['Average Reward'] for p in all_performance_data]
    convergences = [p['Convergence'] for p in all_performance_data]
    environment_adaptabilities = [p['Environment Adaptability'] for p in all_performance_data]
    generalization_abilities = [p['Generalization Ability'] for p in all_performance_data]
    computational_efficiencies = [p['Computational Efficiency'] for p in all_performance_data]
    policy_update_frequencies = [p['Policy Update Frequency'] for p in all_performance_data]
    avg_similarities = [p['Average Similarity'] for p in all_performance_data]

    print(f"\nTotal Reward Statistics:")
    print(f"  Mean: {np.mean(total_rewards):.2f}")
    print(f"  Std: {np.std(total_rewards):.2f}")

    print(f"\nAverage Reward Statistics:")
    print(f"  Mean: {np.mean(average_rewards):.4f}")
    print(f"  Std: {np.std(average_rewards):.4f}")

    print(f"\nConvergence Statistics:")
    print(f"  Mean: {np.mean(convergences):.4f}")
    print(f"  Std: {np.std(convergences):.4f}")

    print(f"\nEnvironment Adaptability Statistics:")
    print(f"  Mean: {np.mean(environment_adaptabilities):.4f}")
    print(f"  Std: {np.std(environment_adaptabilities):.4f}")

    print(f"\nGeneralization Ability Statistics:")
    print(f"  Mean: {np.mean(generalization_abilities):.4f}")
    print(f"  Std: {np.std(generalization_abilities):.4f}")

    print(f"\nComputational Efficiency Statistics:")
    print(f"  Mean: {np.mean(computational_efficiencies):.2f}")
    print(f"  Std: {np.std(computational_efficiencies):.2f}")

    print(f"\nPolicy Update Frequency Statistics:")
    print(f"  Mean: {np.mean(policy_update_frequencies):.4f}")
    print(f"  Std: {np.std(policy_update_frequencies):.4f}")

    print("\n" + "=" * 80)
    print(f" {EXPERIMENT_CONFIG['NUM_RUNS']} 次运行全部完成!")
    print("=" * 80)

if __name__ == "__main__":
    main()