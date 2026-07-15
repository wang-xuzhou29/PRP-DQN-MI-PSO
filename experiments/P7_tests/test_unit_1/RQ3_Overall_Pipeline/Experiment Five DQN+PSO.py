import os
import sys
import random
import time
from collections import deque
from datetime import datetime
from typing import Any, Dict, List, Optional, Set, Tuple

import numpy as np
import torch
import torch.nn as nn
import torch.optim as optim
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ========================================
# ========== experiment configuration parameters ==========
# ========================================
NUM_RUNS = 20

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === three-dimensional range settings ===
# Keep the current DQN state range used by the second script. To use a 0-500 range, modify this section only.
LIGHT_MIN = 1
LIGHT_MAX = 300
MOISTURE_MIN = 1
MOISTURE_MAX = 200
TEMP_MIN = 1
TEMP_MAX = 5
BOUNDS = {
    "light": (LIGHT_MIN, LIGHT_MAX),
    "moisture": (MOISTURE_MIN, MOISTURE_MAX),
    "temp": (TEMP_MIN, TEMP_MAX),
}

# === standard PSO parameters ===
PSO_W = 0.7
PSO_C1 = 1.5
PSO_C2 = 1.5
PSO_VMAX_RATIO = 0.2


def normalize_state(state):
    """Normalize the state to the [0, 1] range."""
    return np.array([
        (state[0] - BOUNDS["light"][0]) / (BOUNDS["light"][1] - BOUNDS["light"][0]),
        (state[1] - BOUNDS["moisture"][0]) / (BOUNDS["moisture"][1] - BOUNDS["moisture"][0]),
        (state[2] - BOUNDS["temp"][0]) / (BOUNDS["temp"][1] - BOUNDS["temp"][0]),
    ], dtype=np.float32)


def denormalize_state(normalized_state):
    """Restore the normalized state to the original range."""
    return np.array([
        normalized_state[0] * (BOUNDS["light"][1] - BOUNDS["light"][0]) + BOUNDS["light"][0],
        normalized_state[1] * (BOUNDS["moisture"][1] - BOUNDS["moisture"][0]) + BOUNDS["moisture"][0],
        normalized_state[2] * (BOUNDS["temp"][1] - BOUNDS["temp"][0]) + BOUNDS["temp"][0],
    ], dtype=float)


def generate_random_state():
    """Generate a random state within each dimension range."""
    light = np.random.randint(BOUNDS["light"][0], BOUNDS["light"][1] + 1)
    moisture = np.random.randint(BOUNDS["moisture"][0], BOUNDS["moisture"][1] + 1)
    temp = np.random.randint(BOUNDS["temp"][0], BOUNDS["temp"][1] + 1)
    return np.array([light, moisture, temp], dtype=float)


def clip_state(state):
    """Clip the state to each dimension boundary."""
    return np.array([
        np.clip(state[0], BOUNDS["light"][0], BOUNDS["light"][1]),
        np.clip(state[1], BOUNDS["moisture"][0], BOUNDS["moisture"][1]),
        np.clip(state[2], BOUNDS["temp"][0], BOUNDS["temp"][1]),
    ], dtype=float)


def is_state_valid(state):
    """Check whether the state is within all valid dimension ranges."""
    return (
        BOUNDS["light"][0] <= state[0] <= BOUNDS["light"][1]
        and BOUNDS["moisture"][0] <= state[1] <= BOUNDS["moisture"][1]
        and BOUNDS["temp"][0] <= state[2] <= BOUNDS["temp"][1]
    )


def execute_Tr(x, y, z):
    """
    根据三维点 a=[x, y, z] 触发一组触发规则，返回满足条件的规则编号集合。
    """
    b = {}
    triggered = set()

    # 规则1-10
    if (x > 190 and x < 210 and y > 90 and y < 110) != (x > 190 and x < 210 and y * 2 > 90 and y < 110): b[0] = 1
    if (x > 190 and x < 210 and y > 90 and y < 110) != (x > 190 and x < 210 and y * 4 > 90 and y < 110): b[1] = 2
    if (x > 190 and x < 210 and y > 90 and y < 110) != (x > 190 and x < 210 and y * 8 > 90 and y < 110): b[2] = 3
    if (y > 85 and y < 115 and z > 1.25 and z < 1.45) != (y > 85 and y < 115 and z > 1.25 and z < 15): b[3] = 4
    if (x > 195 and x < 205) != (x * 8 > 195 and x < 205): b[4] = 5
    if (x > 195 and x < 205) != (x > 195 and x * 8 < 205): b[5] = 6
    if (x > 195 and x < 205) != (x > 195 and x < 605): b[6] = 7
    if (y > 95 and y < 105) != (y > 95 and y * 8 < 105): b[7] = 8
    if (z > 1.3 and z < 1.4) != (z > 1.3 and z < 14): b[8] = 9
    if (z > 1 and z < 3.4) != (z > 1 and z < 4.4): b[9] = 10

    # 规则11-20
    if (abs(x - 200) < 8) != (abs(x - 200) < 18): b[10] = 11
    if (abs(y - 100) < 8) != (abs(y ** 2 - 100) < 8): b[11] = 12
    if (abs(z - 1.35) < 0.08) != (abs(z - 1.35) < 1.08): b[12] = 13
    if (abs(z - 1.35) < 0.08) != (abs(z - 1.35) < 2.08): b[13] = 14
    if (abs(z - 1.35) < 1.08) != (abs(z + 1.35) < 1.08): b[14] = 15
    if (x > 188 and x < 212 and y > 88 and y < 112 and z > 1.28 and z < 1.42) != (
            x > 188 and x < 212 and y > 88 and y < 112 and z > 1.28 and z < 9.42): b[15] = 16
    if ((x + y + z * 100) > 420 and (x + y + z * 100) < 460) != ((x + y + z * 100) > 720 and (x + y + z * 100) < 460):
        b[16] = 17
    if ((x + y + z * 100) > 420 and (x + y + z * 100) < 460) != ((x + y + z * 100) > 420 and (x + y + z * 100) < 260):
        b[17] = 18
    if ((x + y + z * 100) > 420 and (x + y + z * 100) < 460) != ((x + y + z * 100) > 420 and (x + y + z * 100) < 360):
        b[18] = 19
    if ((x + y + z * 100) > 420 and (x + y + z * 100) < 460) != ((x + y + z * 100) > 420 and (x + y + z * 100) < 560):
        b[19] = 20

    # 规则21-30
    if (x / y > 1.8 and x / y < 2.2) != (x / y > 1.8 and x / y < 22): b[20] = 21
    if (x / y > 1.8 and x / y < 2.2) != (x / y > 1.8 and x / y < 12.2): b[21] = 22
    if (x / z > 130 and x / z < 170) != (x / z > 130 and x / z < 1170): b[22] = 23
    if (x / z > 130 and x / z < 170) != (x / z > 130 and x / z < 120): b[23] = 24
    if (x / z > 130 and x / z < 170) != (x / z > 130 and x / z < 270): b[24] = 25
    if (y / z > 60 and y / z < 90) != (y / z > 60 and y / z < 910): b[25] = 26
    if (y / z > 60 and y / z < 90) != (y / z > 60 and y / z < 190): b[26] = 27
    if (y / z > 60 and y / z < 90) != (y / z > 60 and y / z < 60): b[27] = 28
    if (x * y > 16000 and x * y < 24000) != (x * y > 16000 and x * y < 2400): b[28] = 29
    if (x * y > 16000 and x * y < 24000) != (x * y > 16000 and x * y < 12400): b[29] = 30

    # 规则31-40
    if (x * y > 16000 and x * y < 24000) != (x * y > 16000 and 22 * y < 24000): b[30] = 31
    if (x * z > 240 and x * z < 320) != (x * z > 240 and x * 22 < 320): b[31] = 32
    if (x * z > 240 and x * z < 320) != (x * z > 240 and 22 * z < 320): b[32] = 33
    if (x * z > 240 and x * z < 320) != (x * z > 240 and x * 99 < 320): b[33] = 34
    if (y * z > 100 and y * z < 170) != (y * z > 100 and y * 22 < 170): b[34] = 35
    if (y * z > 100 and y * z < 170) != (y * z > 100 and 122 * z < 170): b[35] = 36
    if (y * z > 100 and y * z < 170) != (y * z > 100 and y * 88 < 170): b[36] = 37
    if ((x + y) / 2 > 145 and (x + y) / 2 < 155) != ((x + y) / 2 > 145 and (x + y) / 4 < 155): b[37] = 38
    if ((x + y) / 2 > 145 and (x + y) / 2 < 155) != ((x + y) / 2 > 145 and (x + y) / 2 < 455): b[38] = 39
    if ((x + y) / 2 > 145 and (x + y) / 2 < 155) != ((x + y) / 2 > 145 and (x + y) / 2 < 855): b[39] = 40

    # 规则41-50
    if (x - y > 60 and x - y < 120) != (x - y > 60 and x - y < 420): b[40] = 41
    if (x - y > 60 and x - y < 120) != (x - y > 60 and x - y < 620): b[41] = 42
    if (x - y > 60 and x - y < 120) != (x - y > 60 and x - y < 820): b[42] = 43
    if (abs(x / y - 2.0) < 0.3) != (abs(x / y - 2.0) < 1.3): b[43] = 44
    if (abs(x / y - 2.0) < 0.3) != (abs(x / y - 2.0) < 2.3): b[44] = 45
    if (abs(x / y - 2.0) < 0.3) != (abs(x / y - 2.0) < 55.3): b[45] = 46
    if (x / (y + 50) > 1.2 and x / (y + 50) < 1.8) != (x / (y + 50) > 1.2 and x / (y + 50) < 18): b[46] = 47
    if (x / (y + 50) > 1.2 and x / (y + 50) < 1.8) != (x / (y + 50) > 1.2 and x / (y + 50) < 221.8): b[47] = 48
    if (x / (y + 50) > 1.2 and x / (y + 50) < 1.8) != (x / (y + 50) > 1.2 and x / (y + 50) < 118): b[48] = 49
    if (y / (z * 50) > 1.2 and y / (z * 50) < 1.8) != (y / (z * 50) > 1.2 and y / (z * 50) < 1118): b[49] = 50

    # 规则51-60
    if (y / (z * 50) > 1.2 and y / (z * 50) < 1.8) != (y / (z * 50) > 1.2 and y / (z * 50) < 221.8): b[50] = 51
    if (y / (z * 50) > 1.2 and y / (z * 50) < 1.8) != (y / (z * 50) > 1.2 and y / (z * 50) < 122.8): b[51] = 52
    if (z / (x / 200) > 1.2 and z / (x / 200) < 1.8) != (z / (x / 200) > 1.2 and z / (x / 200) < 111.8): b[52] = 53
    if (z / (x / 200) > 1.2 and z / (x / 200) < 1.8) != (z / (x / 200) > 1.2 and z / (x / 200) < 18): b[53] = 54
    if (z / (x / 200) > 1.2 and z / (x / 200) < 1.8) != (z / (x / 200) > 1.2 and z / (x / 200) < 1.811): b[54] = 55
    if ((x * y * z) > 25000 and (x * y * z) < 35000) != ((x * y * z) > 25000 and (x * y * z) < 300): b[55] = 56
    if ((x + y + z) / 3 > 108 and (x + y + z) / 3 < 118) != ((x + y + z) / 3 > 108 and (x + y + z) / 5 < 118): b[56] = 57
    if ((x + y + z) / 3 > 108 and (x + y + z) / 3 < 118) != ((x + y + z) / 3 > 108 and (x + y + z) / 9 < 118): b[57] = 58
    if ((x + y + z) / 3 > 108 and (x + y + z) / 3 < 118) != ((x + y + z) / 3 > 108 and (x + y + z) / 8 < 118): b[58] = 59
    if (x ** 2 + y ** 2 + z ** 2 * 10000 > 50000) != (x ** 2 + y ** 2 + z ** 2 * 10000 > 80000): b[59] = 60

    # 规则61-70
    if (x ** 2 + y ** 2 + z ** 2 * 10000 > 50000) != (x ** 2 + y ** 2 + z ** 2 * 10000 > 99000): b[60] = 61
    if (x ** 2 + y ** 2 + z ** 2 * 10000 > 50000) != (x ** 2 + y ** 2 + z ** 2 * 10000 > 88000): b[61] = 62
    if ((x - 180) * (y - 80) > 200 and (x - 180) * (y - 80) < 1000) != (
            (x - 180) * (y - 80) > 200 and (x - 180) * (y - 80) < 4000): b[62] = 63
    if (abs((x + y) - 300) < 20) != (abs((x + y) - 300) < 120): b[63] = 64
    if (abs((x + y) - 300) < 20) != (abs((x + y) - 300) < 210): b[64] = 65
    if (abs((x + y) - 300) < 20) != (abs((x + y) - 300) < 2990): b[65] = 66
    if (x / 200 > 0.9 and x / 200 < 1.1) != (x / 200 > 0.9 and x / 200 < 11): b[66] = 67
    if (x / 200 > 0.9 and x / 200 < 1.1) != (x / 200 > 0.9 and x / 200 < 8.1): b[67] = 68
    if (x / 200 > 0.9 and x / 200 < 1.1) != (x / 200 > 0.9 and x / 200 < 99.1): b[68] = 69
    if (y / 100 > 0.9 and y / 100 < 1.1) != (y / 100 > 0.9 and y / 100 < 11): b[69] = 70

    # 规则71-80
    if (y / 100 > 0.9 and y / 100 < 1.1) != (y / 100 > 0.9 and y / 10 < 1.1): b[70] = 71
    if (y / 100 > 0.9 and y / 100 < 1.1) != (y / 100 > 0.9 and y / 200 < 1.1): b[71] = 72
    if (x < 175) != (x < 115): b[72] = 73
    if (x < 175) != (x < 125): b[73] = 74
    if (x > 225) != (x > 2215): b[74] = 75
    if (x > 225) != (x > 2299): b[75] = 76
    if (x > 225) != (x > 1225): b[76] = 77
    if (y < 75) != (y < 175): b[77] = 78
    if (y < 75) != (y < 715): b[78] = 79
    if (y < 75) != (y < 751): b[79] = 80

    # 规则81-90
    if (y > 125) != (y > 925): b[80] = 81
    if (y > 125) != (y > 1115): b[81] = 82
    if (y > 125) != (y > 1215): b[82] = 83
    if (z < 1.15) != (z < 3.85): b[83] = 84
    if (z < 1.15) != (z < 4.5): b[84] = 85
    if (z < 2.15) != (z < 0.15): b[85] = 86
    if (z > 2.55) != (z > 1.25): b[86] = 87
    if (z > 1.55) != (z > 7.45): b[87] = 88
    if (z > 1.55) != (z > 165): b[88] = 89
    if (x < 170 or x > 230) != (x < 170 or x > 1230): b[89] = 90

    # 规则91-100
    if (x < 170 or x > 230) != (x < 170 or x > 2130): b[90] = 91
    if (x < 165 and y < 75) != (x < 165 and y * 8 < 75): b[91] = 92
    if (x > 235 and y > 125) != (x > 235 and y ** 9 > 125): b[92] = 93
    if (x > 235 and y > 125) != (x > 235 and y ** 2 > 125): b[93] = 94
    if (x > 235 and y > 125) != (x > 235 and y ** 4 > 125): b[94] = 95
    if (x < 170 and z < 1.15) != (x < 170 and z < 4.15): b[95] = 96
    if (x < 170 and z < 1.15) != (x < 170 and z < 9.15): b[96] = 97
    if (x < 170 and z < 1.15) != (x < 170 and z < 94.15): b[97] = 98
    if (x > 230 and z > 1.55) != (x > 230 and z > 155): b[98] = 99
    if (x > 230 and z > 1.55) != (x > 230 and z > 4.55): b[99] = 100

    # 规则101-110
    if (y > 125 and z > 1.55) != (y > 125 and z * 2 > 1.55): b[100] = 101
    if (y > 125 and z > 1.55) != (y > 125 and z * 6 > 1.55): b[101] = 102
    if (abs(x / y - 2.0) > 0.5) != (abs(x / y - 2.0) > 10.5): b[102] = 103
    if (abs(x / y - 2.0) > 0.5) != (abs(x / y - 2.0) > 120.5): b[103] = 104
    if (abs(x / y - 2.0) > 0.5) != (abs(x / y - 2.0) > 5): b[104] = 105
    if (x * y < 14000 or x * y > 26000) != (x * y < 14000 or x * y > 28000): b[105] = 106
    if (x * y < 14000 or x * y > 26000) != (x * y < 14000 or x * y > 27000): b[106] = 107
    if (x * y < 14000 or x * y > 26000) != (x * y < 14000 or x * y > 30000): b[107] = 108
    if ((x + y + z * 100) < 400 or (x + y + z * 100) > 480) != ((x + y + z * 100) < 400 or (x + y + z * 100) > 4180): b[108] = 109
    if ((x + y + z * 100) < 400 or (x + y + z * 100) > 480) != ((x + y + z * 100) < 400 or (x + y + z * 1100) > 480): b[109] = 110

    # 规则111-120
    if ((x + y + z * 100) < 400 or (x + y + z * 100) > 480) != ((x + y + z * 100) < 400 or (x + y + z * 800) > 480): b[110] = 111
    if (x < 160 or x > 240) != (x < 160 or x > 1240): b[111] = 112
    if (x < 160 or x > 240) != (x < 160 or x > 2140): b[112] = 113
    if (x < 160 or x > 240) != (x < 160 or x > 2410): b[113] = 114
    if (y < 60 or y > 140) != (y < 60 or y > 1410): b[114] = 115
    if (y < 60 or y > 140) != (y < 60 or y > 1420): b[115] = 116
    if (y < 60 or y > 140) != (y < 60 or y > 1770): b[116] = 117
    if (z > 1.33 and z < 1.37) != (z > 1.33 and z < 137): b[117] = 118
    if (abs(x - 200) < 3 and abs(y - 100) < 3) != (abs(x - 200) < 3 and abs(y - 100) < 13): b[118] = 119
    if (abs(x - 200) < 3 and abs(y - 100) < 3) != (abs(x - 200) < 3 and abs(y - 100) < 322): b[119] = 120

    # 规则121-130
    if (abs(x - 200) < 3 and abs(y - 100) < 3) != (abs(x - 200) < 3 and abs(y - 100) < 31): b[120] = 121
    if (abs(z - 1.35) < 0.03) != (abs(z - 1.35) < 4.03): b[121] = 122
    if (abs(z - 1.35) < 0.03) != (abs(z - 1.35) < 8.03): b[122] = 123
    if (abs(z - 1.35) < 2.03) != (abs(z - 1.35) < 1.3): b[123] = 124
    if (x / y > 1.95 and x / y < 2.05) != (x / y > 125 and x / y < 2.05): b[124] = 125
    if (x / y > 1.95 and x / y < 2.05) != (x / y > 1.95 and x / y < 205): b[125] = 126
    if (x / y > 1.95 and x / y < 2.05) != (x / y > 1.95 and x / y < 2105): b[126] = 127
    if (x * y > 19000 and x * y < 21000) != (x * y > 19000 and x * y < 1000): b[127] = 128
    if (x * y > 19000 and x * y < 21000) != (x * y > 1900 and x * y < 21000): b[128] = 129
    if (x * y > 19000 and x * y < 21000) != (x * y > 19000 and x * y < 21090): b[129] = 130

    # 规则131-140
    if ((x + y + z * 100) > 430 and (x + y + z * 100) < 440) != ((x + y + z * 100) > 430 and (x + y + z * 100) < 4140): b[130] = 131
    if ((x + y + z * 100) > 430 and (x + y + z * 100) < 440) != ((x + y + z * 100) > 430 and (x + y + z * 100) < 1440): b[131] = 132
    if ((x + y + z * 100) > 430 and (x + y + z * 100) < 440) != ((x + y + z * 100) > 430 and (x + y + z * 100) < 2440): b[132] = 133
    if ((x * y * z) > 29000 and (x * y * z) < 31000) != ((x * y * z) > 29000 and (x * y * z) < 41000): b[133] = 134
    if ((x * y * z) > 29000 and (x * y * z) < 31000) != ((x * y * z) > 29000 and (x * y * z) < 38000): b[134] = 135
    if ((x * y * z) > 29000 and (x * y * z) < 31000) != ((x * y * z) > 29000 and (x * y * z) < 39000): b[135] = 136
    if (((x - 200) ** 2 + (y - 100) ** 2) < 25) != (((x - 200) ** 2 + (y - 100) ** 2) < 125): b[136] = 137
    if (((x - 200) ** 2 + (y - 100) ** 2) < 25) != (((x - 200) ** 2 + (y - 100) ** 2) < 215): b[137] = 138
    if (((x - 200) ** 2 + (y - 100) ** 2) < 25) != (((x - 200) ** 2 + (y - 100) ** 2) < 251): b[138] = 139
    if (abs((x + y) / 2 - 150) < 2) != (abs((x + y) / 2 - 150) < 12): b[139] = 140

    # 规则141-150
    if (abs((x + y) / 2 - 150) < 2) != (abs((x + y) / 2 - 150) < 21): b[140] = 141
    if (abs((x + y) / 2 - 150) < 2) != (abs((x + y) / 2 - 150) < 9): b[141] = 142
    if (x / (z * 100) > 1.45 and x / (z * 100) < 1.55) != (x / (z * 100) > 1.45 and x / (z * 100) < 11.55): b[142] = 143
    if (y / (z * 50) > 1.45 and y / (z * 50) < 1.55) != (y / (z * 50) > 1.45 and y / (z * 50) < 55): b[143] = 144
    if ((x - 180) / 40 > 0.45 and (x - 180) / 40 < 0.55) != ((x - 180) / 40 > 0.45 and (x - 180) / 4 < 0.55): b[144] = 145
    if ((x - 180) / 40 > 0.45 and (x - 180) / 40 < 0.55) != ((x - 180) / 40 > 0.45 and (x - 180) / 90 < 0.55): b[145] = 146
    if ((x - 180) / 40 > 0.45 and (x - 180) / 40 < 0.55) != ((x - 180) / 40 > 0.45 and (x - 180) / 80 < 0.55): b[146] = 147
    if ((y - 80) / 40 > 0.45 and (y - 80) / 40 < 0.55) != ((y - 80) / 40 > 0.45 and (y - 80) / 40 < 22.55): b[147] = 148
    if ((y - 80) / 40 > 0.45 and (y - 80) / 40 < 0.55) != ((y - 80) / 40 > 0.45 and (y - 80) / 40 < 2.55): b[148] = 149
    if ((y - 80) / 40 > 0.45 and (y - 80) / 40 < 0.55) != ((y - 80) / 40 > 0.45 and (y - 80) / 40 < 6.55): b[149] = 150

    # 规则151-160
    if ((z - 1.2) / 0.3 > 0.45 and (z - 1.2) / 0.3 < 0.55) != ((z - 1.2) / 0.3 > 0.45 and (z - 1.2) / 3 < 0.55): b[150] = 151
    if ((z - 1.2) / 0.3 > 0.45 and (z - 1.2) / 0.3 < 0.55) != ((z - 1.2) / 0.3 > 0.45 and (z - 1.2) / 0.3 < 5.55): b[151] = 152
    if ((z - 1.2) / 0.3 > 0.45 and (z - 1.2) / 0.3 < 0.55) != ((z - 1.2) / 0.3 > 0.45 and (z - 1.2) / 3 < 0.55): b[152] = 153
    if (min(x - 180, 220 - x) > 15) != (min(x - 180, 220 - x) > 152): b[153] = 154
    if (min(x - 180, 220 - x) > 15) != (min(x - 180, 220 - x) > 115): b[154] = 155
    if (min(x - 180, 220 - x) > 15) != (min(x - 180, 220 - x) > 5): b[155] = 156
    if (min(y - 80, 120 - y) > 15) != (min(y - 80, 120 + y) > 15): b[156] = 157
    if (min(y - 80, 120 - y) > 15) != (min(y - 80, 120 * y) > 15): b[157] = 158
    if (min(y - 80, 120 - y) > 15) != (min(y - 80, 120 - y) > 95): b[158] = 159
    if (min(y - 80, 120 - y) > 15) != (min(y - 80, 120 - y) > 35): b[159] = 160

    # 规则161-163
    if (abs(x * y - 20000) < 500) != (abs(x * y - 20000) < 1500): b[160] = 161
    if (abs(x * y - 20000) < 500) != (abs(x * y - 20000) < 5100): b[161] = 162
    if (abs(x * y - 20000) < 500) != (abs(x * y - 20000) < 5010): b[162] = 163

    # 将触发的规则添加到结果集合中
    for key, value in b.items():
        triggered.add(value)

    return triggered


target_paths = [
    {7, 8, 12, 13, 14, 15, 20, 21, 22, 23, 25, 26, 27, 31, 32, 34, 35, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 50, 51,
     52, 56, 57, 58, 59, 61, 63, 64, 65, 66, 67, 68, 69, 71, 75, 76, 77, 78, 79, 80, 86, 90, 91, 93, 94, 95, 103, 104,
     105, 106, 107, 108, 110, 111, 112, 113, 114, 122, 123, 126, 127, 131, 132, 133, 143, 144, 148, 149, 150, 159, 160},

    {7, 8, 12, 13, 14, 15, 17, 18, 19, 21, 22, 23, 25, 26, 27, 31, 32, 34, 35, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46,
     50, 51, 52, 56, 57, 58, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 75, 76, 77, 78, 79, 80, 86, 90, 91, 93, 94, 95,
     103, 104, 105, 106, 107, 108, 110, 111, 112, 113, 114, 122, 123, 126, 127, 131, 132, 133, 143, 144, 159, 160},

    {7, 13, 14, 15, 20, 21, 22, 23, 26, 27, 31, 32, 34, 35, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51,
     52, 56, 57, 58, 59, 63, 64, 65, 66, 67, 68, 69, 70, 72, 75, 76, 77, 78, 79, 80, 86, 90, 91, 93, 94, 95, 103, 104,
     105, 109, 112, 113, 114, 122, 123, 126, 127, 131, 132, 133, 134, 135, 136, 143, 144, 148, 149, 150, 157, 158},

    {4, 7, 8, 9, 12, 13, 14, 15, 21, 22, 24, 31, 33, 38, 39, 40, 41, 42, 43, 44, 45, 46, 57, 58, 59, 63, 64, 65, 66, 67,
     68, 69, 71, 75, 76, 77, 78, 79, 80, 84, 85, 86, 87, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 106, 107,
     108, 109, 112, 113, 114, 118, 122, 123, 126, 127, 131, 132, 133, 148, 149, 150, 151, 152, 153, 159, 160},

    {7, 9, 13, 14, 15, 20, 21, 22, 29, 30, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 64, 65, 66,
     67, 68, 69, 75, 76, 77, 78, 79, 80, 84, 85, 86, 87, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 109, 112,
     113, 114, 118, 122, 123, 126, 127, 128, 131, 132, 133, 134, 136, 141, 151, 152, 153, 161, 162, 163},

    {7, 9, 13, 14, 15, 20, 21, 22, 24, 29, 30, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 47, 48, 49, 57, 58, 59, 64,
     65, 66, 67, 68, 69, 75, 76, 77, 84, 85, 86, 87, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 109, 112, 113,
     114, 118, 122, 123, 126, 127, 129, 131, 132, 133, 134, 135, 136, 151, 152, 153, 161, 162, 163},

    {7, 9, 13, 14, 15, 20, 21, 22, 24, 29, 30, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 64, 65,
     66, 67, 68, 69, 75, 76, 77, 78, 79, 80, 84, 85, 86, 87, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 109,
     112, 113, 114, 118, 122, 123, 126, 127, 130, 131, 132, 133, 151, 152, 153, 161, 162, 163},

    {7, 9, 13, 14, 15, 20, 21, 22, 29, 30, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 67, 68, 69,
     75, 76, 77, 78, 79, 80, 84, 85, 86, 87, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 109, 118, 122, 123,
     126, 127, 129, 131, 132, 133, 134, 135, 136, 140, 141, 142, 151, 152, 153, 161, 162, 163},

    {7, 13, 14, 15, 20, 23, 25, 26, 27, 31, 32, 34, 35, 37, 38, 39, 40, 44, 45, 46, 50, 51, 52, 56, 57, 58, 59, 61, 63,
     64, 65, 66, 67, 68, 69, 70, 72, 75, 76, 77, 78, 79, 80, 81, 82, 83, 86, 90, 91, 101, 102, 109, 112, 113, 114, 115,
     116, 117, 122, 123, 131, 132, 133, 134, 135, 136, 143, 144, 148, 149, 150, 157, 158},

    {7, 11, 13, 14, 15, 17, 18, 19, 23, 25, 26, 27, 31, 35, 37, 38, 39, 40, 44, 45, 46, 50, 51, 52, 56, 57, 58, 59, 60,
     61, 62, 63, 64, 65, 66, 70, 72, 78, 79, 80, 81, 82, 83, 86, 101, 102, 103, 104, 105, 110, 111, 115, 116, 117, 122,
     123, 131, 132, 133, 134, 135, 136, 143, 144, 146, 147, 148, 149, 150, 156, 157, 158},

    {5, 9, 13, 14, 15, 20, 28, 29, 30, 32, 34, 44, 45, 46, 53, 54, 60, 61, 62, 64, 65, 66, 70, 72, 73, 78, 79, 80, 81,
     82, 83, 84, 85, 86, 87, 88, 89, 96, 97, 98, 103, 104, 105, 110, 111, 115, 116, 117, 118, 122, 123, 128, 131, 132,
     133, 134, 136, 140, 141, 144, 148, 149, 150, 151, 152, 153, 157, 158, 161, 162, 163},

    {5, 9, 13, 14, 15, 20, 28, 29, 30, 32, 34, 44, 45, 46, 53, 54, 60, 61, 62, 64, 65, 66, 70, 72, 73, 74, 78, 79, 80,
     81, 82, 83, 84, 85, 86, 87, 88, 89, 96, 97, 98, 103, 104, 105, 110, 111, 115, 116, 117, 118, 122, 123, 128, 131,
     132, 133, 134, 136, 140, 141, 148, 149, 150, 151, 152, 153, 157, 158, 161, 162, 163},

    {4, 7, 8, 9, 12, 14, 21, 22, 31, 33, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 53, 54, 57, 58, 59, 63, 64, 65,
     66, 67, 68, 69, 71, 75, 76, 77, 78, 79, 80, 84, 85, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 106, 107,
     108, 109, 112, 113, 114, 118, 122, 123, 124, 126, 127, 131, 132, 133, 159, 160},

    {6, 13, 14, 15, 17, 18, 19, 23, 25, 26, 27, 31, 35, 37, 38, 39, 40, 44, 45, 46, 50, 51, 52, 56, 57, 58, 59, 60, 61,
     62, 63, 64, 65, 66, 70, 72, 78, 79, 80, 81, 82, 83, 86, 101, 102, 103, 104, 105, 110, 111, 115, 116, 117, 120, 122,
     123, 131, 132, 133, 143, 144, 146, 147, 148, 149, 150, 154, 155, 157, 158},

    {6, 13, 14, 15, 17, 18, 19, 23, 25, 26, 27, 31, 35, 37, 38, 39, 40, 44, 45, 46, 50, 51, 52, 56, 60, 61, 62, 63, 64,
     65, 66, 70, 72, 78, 79, 80, 81, 82, 83, 86, 101, 102, 103, 104, 105, 108, 110, 111, 115, 116, 117, 120, 122, 123,
     131, 132, 133, 141, 143, 144, 145, 148, 149, 150, 154, 155, 157, 158},

    {4, 7, 9, 10, 12, 21, 22, 31, 33, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 53, 54, 57, 58, 59, 63, 64, 65,
     66, 67, 68, 69, 71, 75, 76, 77, 78, 79, 80, 85, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 109, 112, 113,
     114, 118, 122, 123, 126, 127, 131, 132, 133, 148, 149, 150, 157, 158},

    {7, 13, 14, 15, 20, 23, 26, 27, 31, 32, 34, 35, 37, 38, 39, 40, 41, 42, 43, 50, 51, 52, 57, 58, 59, 65, 66, 67, 68,
     69, 70, 72, 75, 76, 77, 78, 79, 80, 81, 82, 83, 86, 90, 91, 101, 102, 109, 112, 113, 114, 115, 116, 117, 122, 123,
     125, 131, 132, 133, 134, 143, 144, 148, 149, 150, 157, 158},

    {7, 9, 13, 14, 15, 28, 31, 33, 38, 39, 40, 44, 45, 46, 53, 54, 55, 57, 58, 59, 63, 64, 65, 66, 67, 68, 69, 70, 72,
     78, 79, 80, 81, 82, 83, 84, 85, 86, 87, 88, 89, 103, 104, 105, 109, 115, 116, 117, 118, 122, 123, 131, 132, 133,
     144, 146, 147, 148, 149, 150, 151, 152, 153, 157, 158},

    {1, 2, 3, 7, 9, 13, 14, 15, 20, 21, 22, 29, 30, 33, 35, 36, 37, 41, 42, 43, 44, 45, 46, 53, 54, 56, 61, 78, 79, 80,
     84, 85, 86, 87, 88, 89, 103, 104, 105, 109, 118, 122, 123, 126, 127, 129, 131, 132, 133, 134, 135, 136, 140, 141,
     142, 146, 147, 151, 152, 153, 156, 162, 163},

    {1, 2, 3, 4, 6, 9, 13, 14, 15, 16, 20, 21, 22, 29, 30, 33, 53, 54, 61, 78, 79, 80, 84, 85, 86, 87, 88, 89, 109, 118,
     119, 120, 121, 122, 123, 126, 127, 129, 131, 132, 133, 134, 135, 136, 137, 138, 139, 140, 141, 142, 145, 151, 152,
     153, 154, 155, 162, 163},

    {5, 9, 13, 14, 15, 17, 18, 19, 21, 22, 32, 34, 35, 36, 37, 44, 45, 46, 53, 54, 60, 61, 62, 64, 65, 66, 73, 74, 84,
     85, 86, 87, 88, 89, 92, 96, 97, 98, 110, 111, 118, 122, 123, 126, 127, 129, 151, 152, 153},
]


def jaccard_similarity(set1: Set[int], set2: Set[int]) -> float:
    """If set1 covers target set2, the match score is treated as 1."""
    if set2.issubset(set1):
        return 1.0
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    return intersection / union if union != 0 else 0.0


def compute_reward(state, target_path, triggered):
    """Compute the DQN reward."""
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward


class SimpleReplayBuffer:
    def __init__(self, capacity=5000):
        self.buffer = deque(maxlen=capacity)

    def append(self, experience):
        self.buffer.append(experience)

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return random.sample(self.buffer, len(self.buffer))
        return random.sample(self.buffer, batch_size)

    def __len__(self):
        return len(self.buffer)


class SimpleDQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super().__init__()
        self.fc1 = nn.Linear(state_dim, 64)
        self.fc2 = nn.Linear(64, 32)
        self.fc3 = nn.Linear(32, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)


class SimpleDQNAgent:
    def __init__(self, state_dim, action_dim, learning_rate=0.001):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.epsilon = 0.3
        self.gamma = 0.99

        self.model = SimpleDQN(state_dim, action_dim).to(device)
        self.target_model = SimpleDQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())
        self.replay_buffer = SimpleReplayBuffer(capacity=5000)

    def decode_action(self, action_idx):
        """Decode the action."""
        light_deltas = [int((LIGHT_MAX - LIGHT_MIN) * p) for p in [0.05, 0.1, 0.2, 0.5, 0.7, -0.05, -0.1, -0.2, -0.5, -0.7]]
        moisture_deltas = [int((MOISTURE_MAX - MOISTURE_MIN) * p) for p in [0.05, 0.1, 0.2, 0.5, 0.7, -0.05, -0.1, -0.2, -0.5, -0.7]]
        temp_deltas = [int((TEMP_MAX - TEMP_MIN) * p) for p in [0.05, 0.1, 0.2, 0.5, 0.7, -0.05, -0.1, -0.2, -0.5, -0.7]]

        dim = action_idx // 10
        delta_idx = action_idx % 10

        if dim == 0:
            return (light_deltas[delta_idx], 0, 0)
        if dim == 1:
            return (0, moisture_deltas[delta_idx], 0)
        return (0, 0, temp_deltas[delta_idx])

    def store_transition(self, state, action, reward, next_state, done):
        self.replay_buffer.append((state, action, reward, next_state, done))

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return

        batch = self.replay_buffer.sample(batch_size)
        states, actions, rewards, next_states, dones = zip(*batch)

        states = torch.tensor(np.array(states), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array(next_states), dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))

        loss = nn.MSELoss()(current_q_values, target_q_values)

        self.optimizer.zero_grad()
        loss.backward()
        self.optimizer.step()

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())

    def get_best_samples(self, target_path, num_samples=20):
        """Screen DQN candidate samples from the replay buffer. This stage is counted separately in T_sample_screening."""
        if len(self.replay_buffer) == 0:
            return []

        samples_with_scores = []
        for state_norm, _, _, _, _ in self.replay_buffer.buffer:
            state_original = denormalize_state(state_norm)
            triggered = execute_Tr(state_original)
            reward = compute_reward(state_original, target_path, triggered)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_scores.append((state_original, reward, sim, triggered))

        samples_with_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_scores[:num_samples]


def train_dqn_for_path(path_idx, target_path, num_samples=200):
    """Train a standard DQN for a single path."""
    print(f"  Start training path {path_idx + 1} DQN model...")
    agent = SimpleDQNAgent(state_dim=3, action_dim=30)

    random_states = [normalize_state(generate_random_state()) for _ in range(num_samples)]

    STEPS_PER_SAMPLE = 3
    EPOCHS = 5
    BATCH_SIZE = 32

    step_count = 0
    for _ in range(EPOCHS):
        for state_norm in random_states:
            state_norm = tuple(state_norm)
            state_original = denormalize_state(state_norm)

            for step in range(STEPS_PER_SAMPLE):
                legal_actions = []
                for a in range(agent.action_dim):
                    dx, dy, dz = agent.decode_action(a)
                    next_state_candidate = (state_original[0] + dx, state_original[1] + dy, state_original[2] + dz)
                    if is_state_valid(next_state_candidate):
                        legal_actions.append(a)

                if not legal_actions:
                    break

                if random.random() < agent.epsilon:
                    action = random.choice(legal_actions)
                else:
                    state_tensor = torch.tensor(state_norm, dtype=torch.float32).unsqueeze(0).to(device)
                    with torch.no_grad():
                        q_values = agent.model(state_tensor)[0]
                    action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                dx, dy, dz = agent.decode_action(action)
                next_state_original = (state_original[0] + dx, state_original[1] + dy, state_original[2] + dz)
                next_state_norm = normalize_state(next_state_original)

                triggered = execute_Tr(next_state_original)
                reward = compute_reward(next_state_original, target_path, triggered)
                done = (step == STEPS_PER_SAMPLE - 1)

                agent.store_transition(state_norm, action, reward, next_state_norm, done)

                state_norm = next_state_norm
                state_original = next_state_original
                step_count += 1

                if step_count % 50 == 0 and len(agent.replay_buffer) >= BATCH_SIZE:
                    agent.train(BATCH_SIZE)

                if step_count % 100 == 0:
                    agent.update_target_model()

    print(f"  Path {path_idx + 1} training completed, replay buffer size: {len(agent.replay_buffer)}")
    return agent


class Particle:
    def __init__(self, initial_position=None):
        if initial_position is not None:
            self.position = clip_state(np.array(initial_position, dtype=float))
        else:
            self.position = np.array([
                np.random.uniform(BOUNDS["light"][0], BOUNDS["light"][1]),
                np.random.uniform(BOUNDS["moisture"][0], BOUNDS["moisture"][1]),
                np.random.uniform(BOUNDS["temp"][0], BOUNDS["temp"][1]),
            ], dtype=float)

        spans = np.array([
            BOUNDS["light"][1] - BOUNDS["light"][0],
            BOUNDS["moisture"][1] - BOUNDS["moisture"][0],
            BOUNDS["temp"][1] - BOUNDS["temp"][0],
        ], dtype=float)
        self.velocity = np.random.uniform(-0.1 * spans, 0.1 * spans)
        self.best_position = self.position.copy()
        self.best_fitness = -1.0
        self.fitness = -1.0


class BasicPSO:
    """
    Standard PSO: contains only standard velocity and position updates.
    Mutation coefficients, opposite particles, local-best reset, and other improved PSO mechanisms have been removed.
    """

    def __init__(self, target_path, swarm_size=20, dqn_samples=None):
        self.target_path = target_path
        self.swarm_size = swarm_size
        self.particles = []
        self.global_best_position = None
        self.global_best_fitness = -1.0

        if dqn_samples is not None and len(dqn_samples) > 0:
            num_direct = min(len(dqn_samples), swarm_size)
            for i in range(num_direct):
                state_tuple, _, _, _ = dqn_samples[i]
                self.particles.append(Particle(initial_position=state_tuple))

            while len(self.particles) < swarm_size:
                base_idx = len(self.particles) % len(dqn_samples)
                state_tuple, _, _, _ = dqn_samples[base_idx]
                perturbed = np.array(state_tuple, dtype=float) + np.random.randint(-10, 11, size=3)
                self.particles.append(Particle(initial_position=clip_state(perturbed)))
        else:
            self.particles = [Particle() for _ in range(swarm_size)]

        for particle in self.particles:
            particle.fitness = self.fitness_function(particle.position)
            particle.best_fitness = particle.fitness
            particle.best_position = particle.position.copy()

            if particle.fitness > self.global_best_fitness:
                self.global_best_fitness = particle.fitness
                self.global_best_position = particle.position.copy()

    def fitness_function(self, position):
        try:
            triggered = execute_Tr(position)
            return jaccard_similarity(triggered, self.target_path)
        except Exception:
            return 0.0

    def update(self):
        spans = np.array([
            BOUNDS["light"][1] - BOUNDS["light"][0],
            BOUNDS["moisture"][1] - BOUNDS["moisture"][0],
            BOUNDS["temp"][1] - BOUNDS["temp"][0],
        ], dtype=float)
        max_velocity = PSO_VMAX_RATIO * spans

        for particle in self.particles:
            r1 = np.random.random(3)
            r2 = np.random.random(3)

            particle.velocity = (
                PSO_W * particle.velocity
                + PSO_C1 * r1 * (particle.best_position - particle.position)
                + PSO_C2 * r2 * (self.global_best_position - particle.position)
            )
            particle.velocity = np.clip(particle.velocity, -max_velocity, max_velocity)

            particle.position = clip_state(particle.position + particle.velocity)
            particle.fitness = self.fitness_function(particle.position)

            if particle.fitness > particle.best_fitness:
                particle.best_fitness = particle.fitness
                particle.best_position = particle.position.copy()

            if particle.fitness > self.global_best_fitness:
                self.global_best_fitness = particle.fitness
                self.global_best_position = particle.position.copy()


def _set_header(ws, headers, widths, header_fill, header_font, border, center_align):
    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_row(ws, row_idx, row_data, border, alignment, fill=None):
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.border = border
        cell.alignment = alignment
        if fill is not None:
            cell.fill = fill


def export_time_metrics_to_excel(all_run_results, experiment_total_time=None, filename=None):
    """Export only timing metrics and per-path iteration counts."""
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"DQN_BasicPSO_Time_Metrics_{timestamp}.xlsx"

    output_dir = os.path.join(os.getcwd(), "results")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    num_runs = len(all_run_results)
    num_paths = len(target_paths)

    # Worksheet 1: runtime summary
    ws1 = wb.active
    ws1.title = "Runtime Summary"
    ws1.sheet_view.showGridLines = False
    headers1 = [
        "Run", "Number of Paths", "DQNTraining Total Time(s)", "Sample Screening Total Time(s)", "DQNDirect Check Total Time(s)",
        "PSOInitialization Total Time(s)", "PSOSearch Total Time(s)", "Total Algorithm Time(s)", "Total Iterations", "Average Iterations per Path"
    ]
    widths1 = [12, 10, 18, 18, 22, 18, 18, 16, 14, 18]
    _set_header(ws1, headers1, widths1, header_fill, header_font, border, center_align)

    for run_idx, run_result in enumerate(all_run_results, 1):
        path_results = run_result["paths"]
        total_dqn_train = sum(p["T_DQN_train"] for p in path_results)
        total_screen = sum(p["T_sample_screening"] for p in path_results)
        total_direct_check = sum(p["T_direct_check"] for p in path_results)
        total_pso_init = sum(p["T_PSO_init"] for p in path_results)
        total_pso_search = sum(p["T_PSO_search"] for p in path_results)
        total_iterations = sum(p["iterations"] for p in path_results)

        row_data = [
            f" {run_idx}",
            num_paths,
            round(total_dqn_train, 6),
            round(total_screen, 6),
            round(total_direct_check, 6),
            round(total_pso_init, 6),
            round(total_pso_search, 6),
            round(run_result["T_run_algorithm_total"], 6),
            int(total_iterations),
            round(total_iterations / num_paths, 2),
        ]
        _write_row(ws1, run_idx + 1, row_data, border, center_align, alternate_fill if run_idx % 2 == 0 else None)

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:J{num_runs + 1}"

    # Worksheet 2: path timing and iteration details
    ws2 = wb.create_sheet(title="Path Timing and Iteration Details")
    ws2.sheet_view.showGridLines = False
    headers2 = [
        "Run", "Path ID", "DQN training(s)", "Sample Screening Time(s)", "DQNDirect Check Time(s)",
        "PSOInitialization Time(s)", "PSOSearch Time(s)", "Path Total Algorithm Time(s)", "Iterations"
    ]
    widths2 = [12, 12, 18, 18, 22, 18, 18, 18, 12]
    _set_header(ws2, headers2, widths2, header_fill, header_font, border, center_align)

    row_idx = 2
    for run_idx, run_result in enumerate(all_run_results, 1):
        for path_result in run_result["paths"]:
            row_data = [
                f" {run_idx}",
                f"Path  {path_result['path_idx'] + 1}",
                round(path_result["T_DQN_train"], 6),
                round(path_result["T_sample_screening"], 6),
                round(path_result["T_direct_check"], 6),
                round(path_result["T_PSO_init"], 6),
                round(path_result["T_PSO_search"], 6),
                round(path_result["T_path_algorithm_total"], 6),
                int(path_result["iterations"]),
            ]
            _write_row(ws2, row_idx, row_data, border, center_align, alternate_fill if row_idx % 2 == 0 else None)
            row_idx += 1

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:I{row_idx - 1}"

    # Worksheet 3: path aggregate statistics
    ws3 = wb.create_sheet(title="Path Aggregate Statistics")
    ws3.sheet_view.showGridLines = False
    headers3 = [
        "Path ID", "Average DQN Training Time(s)", "Average Sample Screening Time(s)", "Average PSO Initialization Time(s)",
        "Average PSO Search Time(s)", "Average Path Total Algorithm Time(s)", "Average Iterations", "Minimum Iterations", "Maximum Iterations"
    ]
    widths3 = [12, 22, 22, 22, 22, 22, 16, 14, 14]
    _set_header(ws3, headers3, widths3, header_fill, header_font, border, center_align)

    for path_idx in range(num_paths):
        records = [run_result["paths"][path_idx] for run_result in all_run_results]
        iterations = [r["iterations"] for r in records]
        row_data = [
            f"Path  {path_idx + 1}",
            round(float(np.mean([r["T_DQN_train"] for r in records])), 6),
            round(float(np.mean([r["T_sample_screening"] for r in records])), 6),
            round(float(np.mean([r["T_PSO_init"] for r in records])), 6),
            round(float(np.mean([r["T_PSO_search"] for r in records])), 6),
            round(float(np.mean([r["T_path_algorithm_total"] for r in records])), 6),
            round(float(np.mean(iterations)), 2),
            int(np.min(iterations)),
            int(np.max(iterations)),
        ]
        _write_row(ws3, path_idx + 2, row_data, border, center_align, alternate_fill if (path_idx + 1) % 2 == 0 else None)

    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:I{num_paths + 1}"

    # Worksheet 4: experiment overview
    ws4 = wb.create_sheet(title="Experiment Overview")
    ws4.sheet_view.showGridLines = False
    headers4 = ["Metric", "Value"]
    widths4 = [32, 24]
    _set_header(ws4, headers4, widths4, header_fill, header_font, border, center_align)

    all_path_records = [p for run_result in all_run_results for p in run_result["paths"]]
    all_iterations = [p["iterations"] for p in all_path_records]
    overview_rows = [
        ["Experiment Runs", num_runs],
        ["Number of Target Paths", num_paths],
        ["Total Experiment Wall-clock Time(s)", round(experiment_total_time, 6) if experiment_total_time is not None else ""],
        ["Average Total Algorithm Time per Run(s)", round(float(np.mean([r["T_run_algorithm_total"] for r in all_run_results])), 6)],
        ["Average Total Algorithm Time per Path(s)", round(float(np.mean([p["T_path_algorithm_total"] for p in all_path_records])), 6)],
        ["Average DQN Training Time per Path(s)", round(float(np.mean([p["T_DQN_train"] for p in all_path_records])), 6)],
        ["Average Sample Screening Time per Path(s)", round(float(np.mean([p["T_sample_screening"] for p in all_path_records])), 6)],
        ["Average PSO Search Time per Path(s)", round(float(np.mean([p["T_PSO_search"] for p in all_path_records])), 6)],
        ["Average Iterations per Path", round(float(np.mean(all_iterations)), 2)],
        ["Maximum Iterations per Path", int(np.max(all_iterations))],
        ["Minimum Iterations per Path", int(np.min(all_iterations))],
    ]

    for idx, row_data in enumerate(overview_rows, 2):
        _write_row(ws4, idx, row_data, border, center_align, alternate_fill if idx % 2 == 0 else None)

    wb.save(filepath)
    print(f"\n Timing metrics exported to: {filepath}")
    return filepath


def run_single_experiment(run_num, max_iterations=3000):
    """Run one DQN + standard PSO ."""
    print(f"\n{'=' * 80}")
    print(f"Start run  {run_num}  run: DQN + standard PSO ")
    print(f"{'=' * 80}")

    run_start = time.perf_counter()
    path_results = []

    for path_idx, target_path in enumerate(target_paths):
        print(f"\nRun {run_num} run - Path {path_idx + 1}")
        path_total_start = time.perf_counter()

        # 1) DQN training
        dqn_train_start = time.perf_counter()
        agent = train_dqn_for_path(path_idx, target_path, num_samples=200)
        T_DQN_train = time.perf_counter() - dqn_train_start

        # 2) DQNSample Screening Time
        screening_start = time.perf_counter()
        dqn_samples = agent.get_best_samples(target_path, num_samples=20)
        T_sample_screening = time.perf_counter() - screening_start

        # 3) DQN
        direct_check_start = time.perf_counter()
        direct_solution_found = False
        for state_tuple, reward, sim, triggered in dqn_samples:
            if sim >= 1.0 or target_path.issubset(triggered):
                direct_solution_found = True
                break
        T_direct_check = time.perf_counter() - direct_check_start

        T_PSO_init = 0.0
        T_PSO_search = 0.0
        iterations_used = 0

        if direct_solution_found:
            print(
                f"  Path {path_idx + 1}: DQN | "
                f"DQN training {T_DQN_train:.6f}s | screening {T_sample_screening:.6f}s | iterations {iterations_used}"
            )
        else:
            # 4) standard PSOInitialization Time
            pso_init_start = time.perf_counter()
            pso = BasicPSO(target_path, swarm_size=20, dqn_samples=dqn_samples)
            T_PSO_init = time.perf_counter() - pso_init_start

            # 5) standard PSOSearch Time
            pso_search_start = time.perf_counter()
            iterations_used = max_iterations
            for iteration in range(max_iterations):
                pso.update()
                if pso.global_best_fitness >= 1.0:
                    iterations_used = iteration + 1
                    break
            T_PSO_search = time.perf_counter() - pso_search_start

            print(
                f"  Path {path_idx + 1}: "
                f"DQN training {T_DQN_train:.6f}s | screening {T_sample_screening:.6f}s | "
                f"PSO {T_PSO_init:.6f}s | PSO {T_PSO_search:.6f}s | "
                f"iterations {iterations_used}"
            )

        T_path_algorithm_total = time.perf_counter() - path_total_start

        path_results.append({
            "path_idx": path_idx,
            "T_DQN_train": T_DQN_train,
            "T_sample_screening": T_sample_screening,
            "T_direct_check": T_direct_check,
            "T_PSO_init": T_PSO_init,
            "T_PSO_search": T_PSO_search,
            "T_path_algorithm_total": T_path_algorithm_total,
            "iterations": iterations_used,
        })

    T_run_algorithm_total = time.perf_counter() - run_start
    print(f"\nRun {run_num} runcompleted | Total Algorithm Time {T_run_algorithm_total:.6f}s")
    return {
        "run_num": run_num,
        "paths": path_results,
        "T_run_algorithm_total": T_run_algorithm_total,
    }


def run_multiple_experiments(num_runs):
    print("\n" + "=" * 80)
    print(f"DQN + standard PSO - {num_runs}: Metric")
    print("=" * 80)
    print(f"Number of Target Paths: {len(target_paths)}")
    print("Statistics: DQN training, Sample Screening Time, PSOInitialization Time, PSOSearch Time, Path Iterations")
    print("=" * 80)

    all_run_results = []
    experiment_start = time.perf_counter()

    for run_num in range(1, num_runs + 1):
        run_result = run_single_experiment(run_num)
        all_run_results.append(run_result)

    experiment_total_time = time.perf_counter() - experiment_start
    print(f"\n{'=' * 80}")
    print(f"All {num_runs} runcompleted | Total Experiment Wall-clock Time {experiment_total_time:.6f}s")
    print(f"{'=' * 80}\n")
    return all_run_results, experiment_total_time


if __name__ == "__main__":
    print("=" * 80)
    print("DQN + standard PSO - MetricPath Iterations")
    print("=" * 80)
    print(f"Current configuration: Run = {NUM_RUNS}")
    print(f"Number of Paths: {len(target_paths)}")
    print(f"Device: {device}")
    print("=" * 80)

    if len(sys.argv) > 1:
        try:
            NUM_RUNS = int(sys.argv[1])
            print(f"Read from command line: Run = {NUM_RUNS}")
        except ValueError:
            print(f"Invalid command-line argument, using default number of runs {NUM_RUNS}")

    all_results, total_time = run_multiple_experiments(num_runs=NUM_RUNS)
    export_time_metrics_to_excel(all_results, total_time)
    print("\nProgram completed")
