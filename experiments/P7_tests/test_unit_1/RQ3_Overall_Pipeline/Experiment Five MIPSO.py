import os
import random
import numpy as np
import time
import psutil
import math
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

NUM_RUNS = 20  # 默认实验运行次数

# === 新的三维范围设置 ===
LIGHT_MIN = 1
LIGHT_MAX = 300
MOISTURE_MIN = 1
MOISTURE_MAX = 200
TEMP_MIN = 1
TEMP_MAX = 5

BOUNDS = {
    'light': (LIGHT_MIN, LIGHT_MAX),
    'temp': (TEMP_MIN, TEMP_MAX),
    'moisture': (MOISTURE_MIN, MOISTURE_MAX)
}


def clip_state(state):
    """将状态限制在各维度的边界内"""
    return np.array([
        np.clip(state[0], BOUNDS['light'][0], BOUNDS['light'][1]),
        np.clip(state[1], BOUNDS['temp'][0], BOUNDS['temp'][1]),
        np.clip(state[2], BOUNDS['moisture'][0], BOUNDS['moisture'][1])
    ])


def execute_Tr(position):
    """执行目标函数并返回触发的路径"""
    x = int(np.clip(position[0], BOUNDS['light'][0], BOUNDS['light'][1]))
    temp = int(np.clip(position[1], BOUNDS['temp'][0], BOUNDS['temp'][1]))
    z = int(np.clip(position[2], BOUNDS['moisture'][0], BOUNDS['moisture'][1]))
    return category1_multivariable_control(x, temp, z)


# === 目标函数 ===
def category1_multivariable_control(x, y, z):
    """
    根据三维点 a=[x, y, z] 触发一组触发规则，返回满足条件的规则编号集合。
    """
    b = {}
    triggered = set()

    # 规则1-10
    if (x > 190 and x < 210 and y > 90 and y < 110) != (x > 190 and x < 210 and y * 2 > 90 and y < 110): b[0] = 1
    if (x > 190 and x < 210 and y > 90 and y < 110) != (x > 190 and x < 210 and y * 4 > 90 and y < 110): b[1] = 2
    if (x > 190 and x < 210 and y > 90 and y < 110) != (x > 190 and x < 210 and y * 8 > 90 and y < 110): b[2] = 3
    if (y > 85 and y < 115 and z > 1.25 and z < 1.45) != (y > 85 and y < 115 and z > 1.25 and z < 15): b[3] = 4
    if (x > 195 and x < 205) != (x * 8 > 195 and x < 205): b[4] = 5
    if (x > 195 and x < 205) != (x > 195 and x * 8 < 205): b[5] = 6
    if (x > 195 and x < 205) != (x > 195 and x < 605): b[6] = 7
    if (y > 95 and y < 105) != (y > 95 and y * 8 < 105): b[7] = 8
    if (z > 1.3 and z < 1.4) != (z > 1.3 and z < 14): b[8] = 9
    if (z > 1 and z < 3.4) != (z > 1 and z < 4.4): b[9] = 10

    # 规则11-20
    if (abs(x - 200) < 8) != (abs(x - 200) < 18): b[10] = 11
    if (abs(y - 100) < 8) != (abs(y ** 2 - 100) < 8): b[11] = 12
    if (abs(z - 1.35) < 0.08) != (abs(z - 1.35) < 1.08): b[12] = 13
    if (abs(z - 1.35) < 0.08) != (abs(z - 1.35) < 2.08): b[13] = 14
    if (abs(z - 1.35) < 1.08) != (abs(z + 1.35) < 1.08): b[14] = 15
    if (x > 188 and x < 212 and y > 88 and y < 112 and z > 1.28 and z < 1.42) != (
            x > 188 and x < 212 and y > 88 and y < 112 and z > 1.28 and z < 9.42): b[15] = 16
    if ((x + y + z * 100) > 420 and (x + y + z * 100) < 460) != ((x + y + z * 100) > 720 and (x + y + z * 100) < 460):
        b[16] = 17
    if ((x + y + z * 100) > 420 and (x + y + z * 100) < 460) != ((x + y + z * 100) > 420 and (x + y + z * 100) < 260):
        b[17] = 18
    if ((x + y + z * 100) > 420 and (x + y + z * 100) < 460) != ((x + y + z * 100) > 420 and (x + y + z * 100) < 360):
        b[18] = 19
    if ((x + y + z * 100) > 420 and (x + y + z * 100) < 460) != ((x + y + z * 100) > 420 and (x + y + z * 100) < 560):
        b[19] = 20

    # 规则21-30
    if (x / y > 1.8 and x / y < 2.2) != (x / y > 1.8 and x / y < 22): b[20] = 21
    if (x / y > 1.8 and x / y < 2.2) != (x / y > 1.8 and x / y < 12.2): b[21] = 22
    if (x / z > 130 and x / z < 170) != (x / z > 130 and x / z < 1170): b[22] = 23
    if (x / z > 130 and x / z < 170) != (x / z > 130 and x / z < 120): b[23] = 24
    if (x / z > 130 and x / z < 170) != (x / z > 130 and x / z < 270): b[24] = 25
    if (y / z > 60 and y / z < 90) != (y / z > 60 and y / z < 910): b[25] = 26
    if (y / z > 60 and y / z < 90) != (y / z > 60 and y / z < 190): b[26] = 27
    if (y / z > 60 and y / z < 90) != (y / z > 60 and y / z < 60): b[27] = 28
    if (x * y > 16000 and x * y < 24000) != (x * y > 16000 and x * y < 2400): b[28] = 29
    if (x * y > 16000 and x * y < 24000) != (x * y > 16000 and x * y < 12400): b[29] = 30

    # 规则31-40
    if (x * y > 16000 and x * y < 24000) != (x * y > 16000 and 22 * y < 24000): b[30] = 31
    if (x * z > 240 and x * z < 320) != (x * z > 240 and x * 22 < 320): b[31] = 32
    if (x * z > 240 and x * z < 320) != (x * z > 240 and 22 * z < 320): b[32] = 33
    if (x * z > 240 and x * z < 320) != (x * z > 240 and x * 99 < 320): b[33] = 34
    if (y * z > 100 and y * z < 170) != (y * z > 100 and y * 22 < 170): b[34] = 35
    if (y * z > 100 and y * z < 170) != (y * z > 100 and 122 * z < 170): b[35] = 36
    if (y * z > 100 and y * z < 170) != (y * z > 100 and y * 88 < 170): b[36] = 37
    if ((x + y) / 2 > 145 and (x + y) / 2 < 155) != ((x + y) / 2 > 145 and (x + y) / 4 < 155): b[37] = 38
    if ((x + y) / 2 > 145 and (x + y) / 2 < 155) != ((x + y) / 2 > 145 and (x + y) / 2 < 455): b[38] = 39
    if ((x + y) / 2 > 145 and (x + y) / 2 < 155) != ((x + y) / 2 > 145 and (x + y) / 2 < 855): b[39] = 40

    # 规则41-50
    if (x - y > 60 and x - y < 120) != (x - y > 60 and x - y < 420): b[40] = 41
    if (x - y > 60 and x - y < 120) != (x - y > 60 and x - y < 620): b[41] = 42
    if (x - y > 60 and x - y < 120) != (x - y > 60 and x - y < 820): b[42] = 43
    if (abs(x / y - 2.0) < 0.3) != (abs(x / y - 2.0) < 1.3): b[43] = 44
    if (abs(x / y - 2.0) < 0.3) != (abs(x / y - 2.0) < 2.3): b[44] = 45
    if (abs(x / y - 2.0) < 0.3) != (abs(x / y - 2.0) < 55.3): b[45] = 46
    if (x / (y + 50) > 1.2 and x / (y + 50) < 1.8) != (x / (y + 50) > 1.2 and x / (y + 50) < 18): b[46] = 47
    if (x / (y + 50) > 1.2 and x / (y + 50) < 1.8) != (x / (y + 50) > 1.2 and x / (y + 50) < 221.8): b[47] = 48
    if (x / (y + 50) > 1.2 and x / (y + 50) < 1.8) != (x / (y + 50) > 1.2 and x / (y + 50) < 118): b[48] = 49
    if (y / (z * 50) > 1.2 and y / (z * 50) < 1.8) != (y / (z * 50) > 1.2 and y / (z * 50) < 1118): b[49] = 50

    # 规则51-60
    if (y / (z * 50) > 1.2 and y / (z * 50) < 1.8) != (y / (z * 50) > 1.2 and y / (z * 50) < 221.8): b[50] = 51
    if (y / (z * 50) > 1.2 and y / (z * 50) < 1.8) != (y / (z * 50) > 1.2 and y / (z * 50) < 122.8): b[51] = 52
    if (z / (x / 200) > 1.2 and z / (x / 200) < 1.8) != (z / (x / 200) > 1.2 and z / (x / 200) < 111.8): b[52] = 53
    if (z / (x / 200) > 1.2 and z / (x / 200) < 1.8) != (z / (x / 200) > 1.2 and z / (x / 200) < 18): b[53] = 54
    if (z / (x / 200) > 1.2 and z / (x / 200) < 1.8) != (z / (x / 200) > 1.2 and z / (x / 200) < 1.811): b[54] = 55
    if ((x * y * z) > 25000 and (x * y * z) < 35000) != ((x * y * z) > 25000 and (x * y * z) < 300): b[55] = 56
    if ((x + y + z) / 3 > 108 and (x + y + z) / 3 < 118) != ((x + y + z) / 3 > 108 and (x + y + z) / 5 < 118): b[56] = 57
    if ((x + y + z) / 3 > 108 and (x + y + z) / 3 < 118) != ((x + y + z) / 3 > 108 and (x + y + z) / 9 < 118): b[57] = 58
    if ((x + y + z) / 3 > 108 and (x + y + z) / 3 < 118) != ((x + y + z) / 3 > 108 and (x + y + z) / 8 < 118): b[58] = 59
    if (x ** 2 + y ** 2 + z ** 2 * 10000 > 50000) != (x ** 2 + y ** 2 + z ** 2 * 10000 > 80000): b[59] = 60

    # 规则61-70
    if (x ** 2 + y ** 2 + z ** 2 * 10000 > 50000) != (x ** 2 + y ** 2 + z ** 2 * 10000 > 99000): b[60] = 61
    if (x ** 2 + y ** 2 + z ** 2 * 10000 > 50000) != (x ** 2 + y ** 2 + z ** 2 * 10000 > 88000): b[61] = 62
    if ((x - 180) * (y - 80) > 200 and (x - 180) * (y - 80) < 1000) != (
            (x - 180) * (y - 80) > 200 and (x - 180) * (y - 80) < 4000): b[62] = 63
    if (abs((x + y) - 300) < 20) != (abs((x + y) - 300) < 120): b[63] = 64
    if (abs((x + y) - 300) < 20) != (abs((x + y) - 300) < 210): b[64] = 65
    if (abs((x + y) - 300) < 20) != (abs((x + y) - 300) < 2990): b[65] = 66
    if (x / 200 > 0.9 and x / 200 < 1.1) != (x / 200 > 0.9 and x / 200 < 11): b[66] = 67
    if (x / 200 > 0.9 and x / 200 < 1.1) != (x / 200 > 0.9 and x / 200 < 8.1): b[67] = 68
    if (x / 200 > 0.9 and x / 200 < 1.1) != (x / 200 > 0.9 and x / 200 < 99.1): b[68] = 69
    if (y / 100 > 0.9 and y / 100 < 1.1) != (y / 100 > 0.9 and y / 100 < 11): b[69] = 70

    # 规则71-80
    if (y / 100 > 0.9 and y / 100 < 1.1) != (y / 100 > 0.9 and y / 10 < 1.1): b[70] = 71
    if (y / 100 > 0.9 and y / 100 < 1.1) != (y / 100 > 0.9 and y / 200 < 1.1): b[71] = 72
    if (x < 175) != (x < 115): b[72] = 73
    if (x < 175) != (x < 125): b[73] = 74
    if (x > 225) != (x > 2215): b[74] = 75
    if (x > 225) != (x > 2299): b[75] = 76
    if (x > 225) != (x > 1225): b[76] = 77
    if (y < 75) != (y < 175): b[77] = 78
    if (y < 75) != (y < 715): b[78] = 79
    if (y < 75) != (y < 751): b[79] = 80

    # 规则81-90
    if (y > 125) != (y > 925): b[80] = 81
    if (y > 125) != (y > 1115): b[81] = 82
    if (y > 125) != (y > 1215): b[82] = 83
    if (z < 1.15) != (z < 3.85): b[83] = 84
    if (z < 1.15) != (z < 4.5): b[84] = 85
    if (z < 2.15) != (z < 0.15): b[85] = 86
    if (z > 2.55) != (z > 1.25): b[86] = 87
    if (z > 1.55) != (z > 7.45): b[87] = 88
    if (z > 1.55) != (z > 165): b[88] = 89
    if (x < 170 or x > 230) != (x < 170 or x > 1230): b[89] = 90

    # 规则91-100
    if (x < 170 or x > 230) != (x < 170 or x > 2130): b[90] = 91
    if (x < 165 and y < 75) != (x < 165 and y * 8 < 75): b[91] = 92
    if (x > 235 and y > 125) != (x > 235 and y ** 9 > 125): b[92] = 93
    if (x > 235 and y > 125) != (x > 235 and y ** 2 > 125): b[93] = 94
    if (x > 235 and y > 125) != (x > 235 and y ** 4 > 125): b[94] = 95
    if (x < 170 and z < 1.15) != (x < 170 and z < 4.15): b[95] = 96
    if (x < 170 and z < 1.15) != (x < 170 and z < 9.15): b[96] = 97
    if (x < 170 and z < 1.15) != (x < 170 and z < 94.15): b[97] = 98
    if (x > 230 and z > 1.55) != (x > 230 and z > 155): b[98] = 99
    if (x > 230 and z > 1.55) != (x > 230 and z > 4.55): b[99] = 100

    # 规则101-110
    if (y > 125 and z > 1.55) != (y > 125 and z * 2 > 1.55): b[100] = 101
    if (y > 125 and z > 1.55) != (y > 125 and z * 6 > 1.55): b[101] = 102
    if (abs(x / y - 2.0) > 0.5) != (abs(x / y - 2.0) > 10.5): b[102] = 103
    if (abs(x / y - 2.0) > 0.5) != (abs(x / y - 2.0) > 120.5): b[103] = 104
    if (abs(x / y - 2.0) > 0.5) != (abs(x / y - 2.0) > 5): b[104] = 105
    if (x * y < 14000 or x * y > 26000) != (x * y < 14000 or x * y > 28000): b[105] = 106
    if (x * y < 14000 or x * y > 26000) != (x * y < 14000 or x * y > 27000): b[106] = 107
    if (x * y < 14000 or x * y > 26000) != (x * y < 14000 or x * y > 30000): b[107] = 108
    if ((x + y + z * 100) < 400 or (x + y + z * 100) > 480) != ((x + y + z * 100) < 400 or (x + y + z * 100) > 4180): b[108] = 109
    if ((x + y + z * 100) < 400 or (x + y + z * 100) > 480) != ((x + y + z * 100) < 400 or (x + y + z * 1100) > 480): b[109] = 110

    # 规则111-120
    if ((x + y + z * 100) < 400 or (x + y + z * 100) > 480) != ((x + y + z * 100) < 400 or (x + y + z * 800) > 480): b[110] = 111
    if (x < 160 or x > 240) != (x < 160 or x > 1240): b[111] = 112
    if (x < 160 or x > 240) != (x < 160 or x > 2140): b[112] = 113
    if (x < 160 or x > 240) != (x < 160 or x > 2410): b[113] = 114
    if (y < 60 or y > 140) != (y < 60 or y > 1410): b[114] = 115
    if (y < 60 or y > 140) != (y < 60 or y > 1420): b[115] = 116
    if (y < 60 or y > 140) != (y < 60 or y > 1770): b[116] = 117
    if (z > 1.33 and z < 1.37) != (z > 1.33 and z < 137): b[117] = 118
    if (abs(x - 200) < 3 and abs(y - 100) < 3) != (abs(x - 200) < 3 and abs(y - 100) < 13): b[118] = 119
    if (abs(x - 200) < 3 and abs(y - 100) < 3) != (abs(x - 200) < 3 and abs(y - 100) < 322): b[119] = 120

    # 规则121-130
    if (abs(x - 200) < 3 and abs(y - 100) < 3) != (abs(x - 200) < 3 and abs(y - 100) < 31): b[120] = 121
    if (abs(z - 1.35) < 0.03) != (abs(z - 1.35) < 4.03): b[121] = 122
    if (abs(z - 1.35) < 0.03) != (abs(z - 1.35) < 8.03): b[122] = 123
    if (abs(z - 1.35) < 2.03) != (abs(z - 1.35) < 1.3): b[123] = 124
    if (x / y > 1.95 and x / y < 2.05) != (x / y > 125 and x / y < 2.05): b[124] = 125
    if (x / y > 1.95 and x / y < 2.05) != (x / y > 1.95 and x / y < 205): b[125] = 126
    if (x / y > 1.95 and x / y < 2.05) != (x / y > 1.95 and x / y < 2105): b[126] = 127
    if (x * y > 19000 and x * y < 21000) != (x * y > 19000 and x * y < 1000): b[127] = 128
    if (x * y > 19000 and x * y < 21000) != (x * y > 1900 and x * y < 21000): b[128] = 129
    if (x * y > 19000 and x * y < 21000) != (x * y > 19000 and x * y < 21090): b[129] = 130

    # 规则131-140
    if ((x + y + z * 100) > 430 and (x + y + z * 100) < 440) != ((x + y + z * 100) > 430 and (x + y + z * 100) < 4140): b[130] = 131
    if ((x + y + z * 100) > 430 and (x + y + z * 100) < 440) != ((x + y + z * 100) > 430 and (x + y + z * 100) < 1440): b[131] = 132
    if ((x + y + z * 100) > 430 and (x + y + z * 100) < 440) != ((x + y + z * 100) > 430 and (x + y + z * 100) < 2440): b[132] = 133
    if ((x * y * z) > 29000 and (x * y * z) < 31000) != ((x * y * z) > 29000 and (x * y * z) < 41000): b[133] = 134
    if ((x * y * z) > 29000 and (x * y * z) < 31000) != ((x * y * z) > 29000 and (x * y * z) < 38000): b[134] = 135
    if ((x * y * z) > 29000 and (x * y * z) < 31000) != ((x * y * z) > 29000 and (x * y * z) < 39000): b[135] = 136
    if (((x - 200) ** 2 + (y - 100) ** 2) < 25) != (((x - 200) ** 2 + (y - 100) ** 2) < 125): b[136] = 137
    if (((x - 200) ** 2 + (y - 100) ** 2) < 25) != (((x - 200) ** 2 + (y - 100) ** 2) < 215): b[137] = 138
    if (((x - 200) ** 2 + (y - 100) ** 2) < 25) != (((x - 200) ** 2 + (y - 100) ** 2) < 251): b[138] = 139
    if (abs((x + y) / 2 - 150) < 2) != (abs((x + y) / 2 - 150) < 12): b[139] = 140

    # 规则141-150
    if (abs((x + y) / 2 - 150) < 2) != (abs((x + y) / 2 - 150) < 21): b[140] = 141
    if (abs((x + y) / 2 - 150) < 2) != (abs((x + y) / 2 - 150) < 9): b[141] = 142
    if (x / (z * 100) > 1.45 and x / (z * 100) < 1.55) != (x / (z * 100) > 1.45 and x / (z * 100) < 11.55): b[142] = 143
    if (y / (z * 50) > 1.45 and y / (z * 50) < 1.55) != (y / (z * 50) > 1.45 and y / (z * 50) < 55): b[143] = 144
    if ((x - 180) / 40 > 0.45 and (x - 180) / 40 < 0.55) != ((x - 180) / 40 > 0.45 and (x - 180) / 4 < 0.55): b[144] = 145
    if ((x - 180) / 40 > 0.45 and (x - 180) / 40 < 0.55) != ((x - 180) / 40 > 0.45 and (x - 180) / 90 < 0.55): b[145] = 146
    if ((x - 180) / 40 > 0.45 and (x - 180) / 40 < 0.55) != ((x - 180) / 40 > 0.45 and (x - 180) / 80 < 0.55): b[146] = 147
    if ((y - 80) / 40 > 0.45 and (y - 80) / 40 < 0.55) != ((y - 80) / 40 > 0.45 and (y - 80) / 40 < 22.55): b[147] = 148
    if ((y - 80) / 40 > 0.45 and (y - 80) / 40 < 0.55) != ((y - 80) / 40 > 0.45 and (y - 80) / 40 < 2.55): b[148] = 149
    if ((y - 80) / 40 > 0.45 and (y - 80) / 40 < 0.55) != ((y - 80) / 40 > 0.45 and (y - 80) / 40 < 6.55): b[149] = 150

    # 规则151-160
    if ((z - 1.2) / 0.3 > 0.45 and (z - 1.2) / 0.3 < 0.55) != ((z - 1.2) / 0.3 > 0.45 and (z - 1.2) / 3 < 0.55): b[150] = 151
    if ((z - 1.2) / 0.3 > 0.45 and (z - 1.2) / 0.3 < 0.55) != ((z - 1.2) / 0.3 > 0.45 and (z - 1.2) / 0.3 < 5.55): b[151] = 152
    if ((z - 1.2) / 0.3 > 0.45 and (z - 1.2) / 0.3 < 0.55) != ((z - 1.2) / 0.3 > 0.45 and (z - 1.2) / 3 < 0.55): b[152] = 153
    if (min(x - 180, 220 - x) > 15) != (min(x - 180, 220 - x) > 152): b[153] = 154
    if (min(x - 180, 220 - x) > 15) != (min(x - 180, 220 - x) > 115): b[154] = 155
    if (min(x - 180, 220 - x) > 15) != (min(x - 180, 220 - x) > 5): b[155] = 156
    if (min(y - 80, 120 - y) > 15) != (min(y - 80, 120 + y) > 15): b[156] = 157
    if (min(y - 80, 120 - y) > 15) != (min(y - 80, 120 * y) > 15): b[157] = 158
    if (min(y - 80, 120 - y) > 15) != (min(y - 80, 120 - y) > 95): b[158] = 159
    if (min(y - 80, 120 - y) > 15) != (min(y - 80, 120 - y) > 35): b[159] = 160

    # 规则161-163
    if (abs(x * y - 20000) < 500) != (abs(x * y - 20000) < 1500): b[160] = 161
    if (abs(x * y - 20000) < 500) != (abs(x * y - 20000) < 5100): b[161] = 162
    if (abs(x * y - 20000) < 500) != (abs(x * y - 20000) < 5010): b[162] = 163

    # 将触发的规则添加到结果集合中
    for key, value in b.items():
        triggered.add(value)

    return triggered


# 目标路径定义
targetPaths = [
    {7, 8, 12, 13, 14, 15, 20, 21, 22, 23, 25, 26, 27, 31, 32, 34, 35, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 50, 51,
     52, 56, 57, 58, 59, 61, 63, 64, 65, 66, 67, 68, 69, 71, 75, 76, 77, 78, 79, 80, 86, 90, 91, 93, 94, 95, 103, 104,
     105, 106, 107, 108, 110, 111, 112, 113, 114, 122, 123, 126, 127, 131, 132, 133, 143, 144, 148, 149, 150, 159, 160},

    {7, 8, 12, 13, 14, 15, 17, 18, 19, 21, 22, 23, 25, 26, 27, 31, 32, 34, 35, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46,
     50, 51, 52, 56, 57, 58, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 75, 76, 77, 78, 79, 80, 86, 90, 91, 93, 94, 95,
     103, 104, 105, 106, 107, 108, 110, 111, 112, 113, 114, 122, 123, 126, 127, 131, 132, 133, 143, 144, 159, 160},

    {7, 13, 14, 15, 20, 21, 22, 23, 26, 27, 31, 32, 34, 35, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51,
     52, 56, 57, 58, 59, 63, 64, 65, 66, 67, 68, 69, 70, 72, 75, 76, 77, 78, 79, 80, 86, 90, 91, 93, 94, 95, 103, 104,
     105, 109, 112, 113, 114, 122, 123, 126, 127, 131, 132, 133, 134, 135, 136, 143, 144, 148, 149, 150, 157, 158},

    {4, 7, 8, 9, 12, 13, 14, 15, 21, 22, 24, 31, 33, 38, 39, 40, 41, 42, 43, 44, 45, 46, 57, 58, 59, 63, 64, 65, 66, 67,
     68, 69, 71, 75, 76, 77, 78, 79, 80, 84, 85, 86, 87, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 106, 107,
     108, 109, 112, 113, 114, 118, 122, 123, 126, 127, 131, 132, 133, 148, 149, 150, 151, 152, 153, 159, 160},

    {7, 9, 13, 14, 15, 20, 21, 22, 29, 30, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 64, 65, 66,
     67, 68, 69, 75, 76, 77, 78, 79, 80, 84, 85, 86, 87, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 109, 112,
     113, 114, 118, 122, 123, 126, 127, 128, 131, 132, 133, 134, 136, 141, 151, 152, 153, 161, 162, 163},

    {7, 9, 13, 14, 15, 20, 21, 22, 24, 29, 30, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 47, 48, 49, 57, 58, 59, 64,
     65, 66, 67, 68, 69, 75, 76, 77, 84, 85, 86, 87, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 109, 112, 113,
     114, 118, 122, 123, 126, 127, 129, 131, 132, 133, 134, 135, 136, 151, 152, 153, 161, 162, 163},

    {7, 9, 13, 14, 15, 20, 21, 22, 24, 29, 30, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 64, 65,
     66, 67, 68, 69, 75, 76, 77, 78, 79, 80, 84, 85, 86, 87, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 109,
     112, 113, 114, 118, 122, 123, 126, 127, 130, 131, 132, 133, 151, 152, 153, 161, 162, 163},

    {7, 9, 13, 14, 15, 20, 21, 22, 29, 30, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 67, 68, 69,
     75, 76, 77, 78, 79, 80, 84, 85, 86, 87, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 109, 118, 122, 123,
     126, 127, 129, 131, 132, 133, 134, 135, 136, 140, 141, 142, 151, 152, 153, 161, 162, 163},

    {7, 13, 14, 15, 20, 23, 25, 26, 27, 31, 32, 34, 35, 37, 38, 39, 40, 44, 45, 46, 50, 51, 52, 56, 57, 58, 59, 61, 63,
     64, 65, 66, 67, 68, 69, 70, 72, 75, 76, 77, 78, 79, 80, 81, 82, 83, 86, 90, 91, 101, 102, 109, 112, 113, 114, 115,
     116, 117, 122, 123, 131, 132, 133, 134, 135, 136, 143, 144, 148, 149, 150, 157, 158},

    {7, 11, 13, 14, 15, 17, 18, 19, 23, 25, 26, 27, 31, 35, 37, 38, 39, 40, 44, 45, 46, 50, 51, 52, 56, 57, 58, 59, 60,
     61, 62, 63, 64, 65, 66, 70, 72, 78, 79, 80, 81, 82, 83, 86, 101, 102, 103, 104, 105, 110, 111, 115, 116, 117, 122,
     123, 131, 132, 133, 134, 135, 136, 143, 144, 146, 147, 148, 149, 150, 156, 157, 158},

    {5, 9, 13, 14, 15, 20, 28, 29, 30, 32, 34, 44, 45, 46, 53, 54, 60, 61, 62, 64, 65, 66, 70, 72, 73, 78, 79, 80, 81,
     82, 83, 84, 85, 86, 87, 88, 89, 96, 97, 98, 103, 104, 105, 110, 111, 115, 116, 117, 118, 122, 123, 128, 131, 132,
     133, 134, 136, 140, 141, 144, 148, 149, 150, 151, 152, 153, 157, 158, 161, 162, 163},

    {5, 9, 13, 14, 15, 20, 28, 29, 30, 32, 34, 44, 45, 46, 53, 54, 60, 61, 62, 64, 65, 66, 70, 72, 73, 74, 78, 79, 80,
     81, 82, 83, 84, 85, 86, 87, 88, 89, 96, 97, 98, 103, 104, 105, 110, 111, 115, 116, 117, 118, 122, 123, 128, 131,
     132, 133, 134, 136, 140, 141, 148, 149, 150, 151, 152, 153, 157, 158, 161, 162, 163},

    {4, 7, 8, 9, 12, 14, 21, 22, 31, 33, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 53, 54, 57, 58, 59, 63, 64, 65,
     66, 67, 68, 69, 71, 75, 76, 77, 78, 79, 80, 84, 85, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 106, 107,
     108, 109, 112, 113, 114, 118, 122, 123, 124, 126, 127, 131, 132, 133, 159, 160},

    {6, 13, 14, 15, 17, 18, 19, 23, 25, 26, 27, 31, 35, 37, 38, 39, 40, 44, 45, 46, 50, 51, 52, 56, 57, 58, 59, 60, 61,
     62, 63, 64, 65, 66, 70, 72, 78, 79, 80, 81, 82, 83, 86, 101, 102, 103, 104, 105, 110, 111, 115, 116, 117, 120, 122,
     123, 131, 132, 133, 143, 144, 146, 147, 148, 149, 150, 154, 155, 157, 158},

    {6, 13, 14, 15, 17, 18, 19, 23, 25, 26, 27, 31, 35, 37, 38, 39, 40, 44, 45, 46, 50, 51, 52, 56, 60, 61, 62, 63, 64,
     65, 66, 70, 72, 78, 79, 80, 81, 82, 83, 86, 101, 102, 103, 104, 105, 108, 110, 111, 115, 116, 117, 120, 122, 123,
     131, 132, 133, 141, 143, 144, 145, 148, 149, 150, 154, 155, 157, 158},

    {4, 7, 9, 10, 12, 21, 22, 31, 33, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 53, 54, 57, 58, 59, 63, 64, 65,
     66, 67, 68, 69, 71, 75, 76, 77, 78, 79, 80, 85, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 109, 112, 113,
     114, 118, 122, 123, 126, 127, 131, 132, 133, 148, 149, 150, 157, 158},

    {7, 13, 14, 15, 20, 23, 26, 27, 31, 32, 34, 35, 37, 38, 39, 40, 41, 42, 43, 50, 51, 52, 57, 58, 59, 65, 66, 67, 68,
     69, 70, 72, 75, 76, 77, 78, 79, 80, 81, 82, 83, 86, 90, 91, 101, 102, 109, 112, 113, 114, 115, 116, 117, 122, 123,
     125, 131, 132, 133, 134, 143, 144, 148, 149, 150, 157, 158},

    {7, 9, 13, 14, 15, 28, 31, 33, 38, 39, 40, 44, 45, 46, 53, 54, 55, 57, 58, 59, 63, 64, 65, 66, 67, 68, 69, 70, 72,
     78, 79, 80, 81, 82, 83, 84, 85, 86, 87, 88, 89, 103, 104, 105, 109, 115, 116, 117, 118, 122, 123, 131, 132, 133,
     144, 146, 147, 148, 149, 150, 151, 152, 153, 157, 158},

    {1, 2, 3, 7, 9, 13, 14, 15, 20, 21, 22, 29, 30, 33, 35, 36, 37, 41, 42, 43, 44, 45, 46, 53, 54, 56, 61, 78, 79, 80,
     84, 85, 86, 87, 88, 89, 103, 104, 105, 109, 118, 122, 123, 126, 127, 129, 131, 132, 133, 134, 135, 136, 140, 141,
     142, 146, 147, 151, 152, 153, 156, 162, 163},

    {1, 2, 3, 4, 6, 9, 13, 14, 15, 16, 20, 21, 22, 29, 30, 33, 53, 54, 61, 78, 79, 80, 84, 85, 86, 87, 88, 89, 109, 118,
     119, 120, 121, 122, 123, 126, 127, 129, 131, 132, 133, 134, 135, 136, 137, 138, 139, 140, 141, 142, 145, 151, 152,
     153, 154, 155, 162, 163},

    {5, 9, 13, 14, 15, 17, 18, 19, 21, 22, 32, 34, 35, 36, 37, 44, 45, 46, 53, 54, 60, 61, 62, 64, 65, 66, 73, 74, 84,
     85, 86, 87, 88, 89, 92, 96, 97, 98, 110, 111, 118, 122, 123, 126, 127, 129, 151, 152, 153},
]


# === 增强版指标收集器 ===
class MetricsCollector:
    def __init__(self):
        self.pso_start_time = None
        self.pso_end_time = None
        self.perfect_solutions_count = 0
        self.total_paths_count = 0
        self.final_fitness_scores = []
        self.pso_convergence_iterations = []
        self.pso_reset_counts = []
        self.path_execution_times = []

    def start_pso_phase(self):
        self.pso_start_time = time.time()

    def end_pso_phase(self):
        self.pso_end_time = time.time()

    def record_pso_result(self, fitness, is_perfect_match, convergence_iter=None, path_id=None, method='MI-PSO',
                          reset_count=0, execution_time=0):
        self.final_fitness_scores.append(fitness)
        self.total_paths_count += 1
        self.pso_reset_counts.append(reset_count)
        self.path_execution_times.append(execution_time)

        if is_perfect_match:
            self.perfect_solutions_count += 1
        if convergence_iter is not None:
            self.pso_convergence_iterations.append(convergence_iter)


# === 粒子类 ===
class Particle:
    def __init__(self):
        self.position = np.array([
            np.random.uniform(BOUNDS['light'][0], BOUNDS['light'][1]),
            np.random.uniform(BOUNDS['temp'][0], BOUNDS['temp'][1]),
            np.random.uniform(BOUNDS['moisture'][0], BOUNDS['moisture'][1])
        ])
        self.velocity = np.array([
            np.random.uniform(-5, 5),
            np.random.uniform(-3, 3),
            np.random.uniform(-5, 5)
        ])
        self.best_position = self.position.copy()
        self.best_fitness = 0
        self.fitness = 0


# === 单独的 MI-PSO 优化器类（已完全修复随机性重评估Bug） ===
class MIPSO:
    def __init__(self, target_path, swarm_size=20, max_iterations=3000, g2=10, th_cv=1.2):
        self.target_path = target_path
        self.swarm_size = swarm_size
        self.max_iterations = max_iterations
        self.g2 = g2
        self.th_cv = th_cv

        # 完全随机初始化
        self.particles = [Particle() for _ in range(swarm_size)]
        self.global_best_position = None
        self.global_best_fitness = 0
        self.global_best_triggered = set()  # 核心修复：记录最高分对应的真实触发路径
        self.reset_count = 0

        # 初始化适应度评估
        for particle in self.particles:
            fit, trig = self.evaluate(particle.position)
            particle.fitness = fit
            particle.best_fitness = fit
            particle.best_position = particle.position.copy()

            if fit > self.global_best_fitness:
                self.global_best_fitness = fit
                self.global_best_position = particle.position.copy()
                self.global_best_triggered = trig

    def evaluate(self, position):
        """同时返回适应度和实际触发的路径"""
        try:
            triggered = execute_Tr(position)
            if self.target_path.issubset(triggered):
                return 1.0, triggered
            intersection = len(triggered & self.target_path)
            union = len(triggered | self.target_path)
            return (intersection / union if union > 0 else 0.0), triggered
        except:
            return 0.0, set()

    def calculate_cv(self, scores):
        """实现公式 25：综合变异系数"""
        N = self.swarm_size
        mean_fit = np.mean(scores)
        if mean_fit == 0: return float('inf')

        var_term = np.sum((scores - mean_fit) ** 2) / N
        adj_diff = sum(abs(scores[2 * i + 1] - scores[2 * i]) for i in range(N // 2))
        adj_term = ((2 / N) * adj_diff) ** 2

        numerator = math.sqrt(var_term + adj_term)
        denominator = 2 * np.sum(scores)

        if denominator == 0: return float('inf')
        return numerator / denominator

    def flip_mutation(self, position, rho=0.1):
        """实现公式 27：变异反转操作"""
        new_pos = np.copy(position)
        new_pos[0] = BOUNDS['light'][0] + BOUNDS['light'][1] - position[0] + rho
        new_pos[1] = BOUNDS['temp'][0] + BOUNDS['temp'][1] - position[1] + rho
        new_pos[2] = BOUNDS['moisture'][0] + BOUNDS['moisture'][1] - position[2] + rho
        return clip_state(new_pos)

    def optimize(self):
        w = 0.7
        c1 = 1.5
        c2 = 1.5

        max_velocity = np.array([
            (BOUNDS['light'][1] - BOUNDS['light'][0]) * 0.2,
            (BOUNDS['temp'][1] - BOUNDS['temp'][0]) * 0.2,
            (BOUNDS['moisture'][1] - BOUNDS['moisture'][0]) * 0.2
        ])

        c = 0
        early_stop = False
        converged_at = self.max_iterations

        for iteration in range(self.max_iterations):
            scores = np.array([p.fitness for p in self.particles])
            c += 1

            # MI-PSO 特有的停滞检测与变异 (对应 Algorithm 4 Line 20-23)
            if c == self.g2:
                cv = self.calculate_cv(scores)
                if cv <= self.th_cv:
                    self.reset_count += 1
                    l = int(round(self.swarm_size * (1 - cv)))
                    l = min(max(l, 1), self.swarm_size)

                    selected_indices = np.random.choice(self.swarm_size, l, replace=False)
                    for idx in selected_indices:
                        new_pos = self.flip_mutation(self.particles[idx].position)
                        new_fit, new_trig = self.evaluate(new_pos)

                        if new_fit > self.particles[idx].fitness:
                            self.particles[idx].position = new_pos
                            self.particles[idx].fitness = new_fit
                            if new_fit > self.particles[idx].best_fitness:
                                self.particles[idx].best_fitness = new_fit
                                self.particles[idx].best_position = new_pos.copy()
                            if new_fit > self.global_best_fitness:
                                self.global_best_fitness = new_fit
                                self.global_best_position = new_pos.copy()
                                self.global_best_triggered = new_trig  # 更新最优路径
                c = 0

                # 标准 PSO 速度与位置更新
            for particle in self.particles:
                r1, r2 = np.random.random(3), np.random.random(3)
                particle.velocity = (w * particle.velocity +
                                     c1 * r1 * (particle.best_position - particle.position) +
                                     c2 * r2 * (self.global_best_position - particle.position))
                particle.velocity = np.clip(particle.velocity, -max_velocity, max_velocity)

                particle.position += particle.velocity
                particle.position = clip_state(particle.position)

                fit, trig = self.evaluate(particle.position)
                particle.fitness = fit

                if fit > particle.best_fitness:
                    particle.best_fitness = fit
                    particle.best_position = particle.position.copy()
                if fit > self.global_best_fitness:
                    self.global_best_fitness = fit
                    self.global_best_position = particle.position.copy()
                    self.global_best_triggered = trig  # 更新最优路径

            if self.global_best_fitness >= 1.0:
                converged_at = iteration + 1
                early_stop = True
                break

        return self.global_best_position, self.global_best_fitness, self.global_best_triggered, converged_at, early_stop, self.reset_count


# === Excel导出函数 ===
def export_multiple_runs_to_excel(all_run_results, all_run_metrics, num_runs, filename=None):
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Isolated_MIPSO_{num_runs}Runs_{timestamp}.xlsx"

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    border = Border(
        left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 工作表1: 运行汇总
    ws1 = wb.active
    ws1.title = "运行汇总"
    ws1.sheet_view.showGridLines = False

    headers = ["运行次数", "成功率", "成功数量", "平均适应度", "平均迭代次数", "运行时间(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, border, center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, (results, run_metrics) in enumerate(zip(all_run_results, all_run_metrics), start=1):
        success_count = sum(1 for r in results if r['perfect_match'])
        success_rate = (success_count / len(targetPaths)) * 100
        avg_fitness = np.mean([r['fitness'] for r in results])

        iterations_list = [r['convergence_iteration'] if r.get('convergence_iteration') is not None else 10000 for r in
                           results]
        avg_iterations = np.mean(iterations_list)
        total_time = run_metrics.pso_end_time - run_metrics.pso_start_time if run_metrics.pso_end_time else 0

        row_data = [f"运行 {run_idx}", f"{success_rate:.1f}%", f"{success_count}/{len(targetPaths)}",
                    f"{avg_fitness:.4f}", f"{avg_iterations:.1f}", f"{total_time:.2f}"]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border, cell.alignment = border, center_align
            if run_idx % 2 == 0: cell.fill = alternate_fill
            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    # 工作表2: 路径统计
    ws2 = wb.create_sheet(title="路径统计")
    ws2.sheet_view.showGridLines = False
    headers2 = ["路径编号", "成功次数", "成功率", "平均适应度", "平均迭代次数", "最小迭代", "最大迭代"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, border, center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(targetPaths)):
        success_count = sum(1 for results in all_run_results if results[path_idx]['perfect_match'])
        success_rate = (success_count / num_runs) * 100
        avg_fitness = np.mean([results[path_idx]['fitness'] for results in all_run_results])

        iterations_list = [results[path_idx]['convergence_iteration'] if results[path_idx].get(
            'convergence_iteration') is not None else 10000 for results in all_run_results]
        avg_iterations = np.mean(iterations_list)

        row_data = [f"路径 {path_idx + 1}", f"{success_count}/{num_runs}", f"{success_rate:.1f}%",
                    f"{avg_fitness:.4f}", f"{avg_iterations:.1f}", np.min(iterations_list), np.max(iterations_list)]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border, cell.alignment = border, center_align
            if (path_idx + 1) % 2 == 0: cell.fill = alternate_fill
            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    # 工作表3: 最佳粒子详情
    ws3 = wb.create_sheet(title="最佳粒子详情")
    ws3.sheet_view.showGridLines = False
    headers3 = ["路径", "运行", "最佳粒子(light,temp,moisture)", "适应度", "迭代次数", "求解方法", "生成路径"]
    col_widths3 = [10, 10, 25, 12, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, border, center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(targetPaths)):
        for run_idx, results in enumerate(all_run_results, start=1):
            result = results[path_idx]
            particle_str = f"({int(result['best_position'][0])}, {int(result['best_position'][1])}, {int(result['best_position'][2])})"
            path_str = str(sorted(list(result['triggered'])))

            convergence_iter = result['convergence_iteration'] if result.get(
                'convergence_iteration') is not None else 10000

            row_data = [f"路径{path_idx + 1}", f"运行{run_idx}", particle_str, f"{result['fitness']:.4f}",
                        convergence_iter if convergence_iter < 10000 else "-", "MI-PSO", path_str]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = left_align if col == 7 else center_align

                if result['fitness'] == 1.0:
                    cell.fill = success_fill
                elif result['fitness'] < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill
            row_idx += 1

    # 保存
    output_dir = os.path.join(os.getcwd(), "results")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"\n✓ 独立 MI-PSO 消融实验结果已导出到: {filepath}")
    return filepath


def run_single_mipso_experiment(run_num):
    print(f"\n{'=' * 50}")
    print(f"开始第 {run_num} 次运行 (独立 MI-PSO)")
    print(f"{'=' * 50}")

    run_metrics = MetricsCollector()
    run_metrics.start_pso_phase()

    pso_results = []

    for i, target_path in enumerate(targetPaths):
        path_start_time = time.time()

        mipso = MIPSO(target_path, swarm_size=20, max_iterations=3000)
        # 直接接收并使用最佳寻优过程中真实触发的 best_trig，坚决不重新 evaluate
        best_pos, best_fit, best_trig, converged_at, early_stop, reset_count = mipso.optimize()

        path_execution_time = time.time() - path_start_time

        # 使用真实的最高分数判断是否完美
        is_perfect = (best_fit >= 1.0)

        pso_results.append({
            'target_path': target_path,
            'best_position': best_pos,
            'fitness': best_fit,
            'triggered': best_trig,
            'perfect_match': is_perfect,
            'method': 'MI-PSO',
            'convergence_iteration': converged_at,
            'early_stopped': early_stop,
            'reset_count': reset_count
        })

        run_metrics.record_pso_result(
            fitness=best_fit,
            is_perfect_match=is_perfect,
            convergence_iter=converged_at if early_stop else None,
            path_id=i + 1,
            method='MI-PSO',
            reset_count=reset_count,
            execution_time=path_execution_time
        )

        status = "✓完美(MI-PSO)" if is_perfect else f"○部分({best_fit:.3f})"
        print(f"  路径{i + 1}: {status} | 变异触发次数: {reset_count} | 耗时 {path_execution_time:.2f}s")

    run_metrics.end_pso_phase()

    success_count = sum(1 for r in pso_results if r['perfect_match'])
    success_rate = (success_count / len(targetPaths)) * 100
    pso_time = run_metrics.pso_end_time - run_metrics.pso_start_time

    print(f"\n第{run_num}次运行完成: 成功率 {success_rate:.1f}% | 耗时 {pso_time:.2f}秒")
    return pso_results, run_metrics


def run_multiple_mipso_experiments(num_runs):
    all_run_results = []
    all_run_metrics = []
    total_start = time.time()

    for run_num in range(1, num_runs + 1):
        results, metrics = run_single_mipso_experiment(run_num)
        all_run_results.append(results)
        all_run_metrics.append(metrics)

    total_time = time.time() - total_start
    print(f"\n{'=' * 50}\n全部{num_runs}次独立 MI-PSO 运行完成! 总耗时: {total_time:.2f}秒\n{'=' * 50}")

    return all_run_results, all_run_metrics


if __name__ == "__main__":
    if len(sys.argv) > 1:
        try:
            NUM_RUNS = int(sys.argv[1])
        except ValueError:
            pass

    print("=" * 70)
    print(" 独立 MI-PSO 消融实验 (无 DQN 介入，完美修复随机性 Bug)")
    print("=" * 70)

    all_run_results, all_run_metrics = run_multiple_mipso_experiments(NUM_RUNS)
    export_multiple_runs_to_excel(all_run_results, all_run_metrics, NUM_RUNS)