import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

def execute_validation_rules(x, y, z):
    """
    根据三维点 a=[x, y, z] 触发一组触发规则，返回满足条件的规则编号集合。
    """
    b = {}
    triggered = set()

    # 规则1-10
    if (x > 190 and x < 210 and y > 90 and y < 110) != (x > 190 and x < 210 and y * 2 > 90 and y < 110): b[0] = 1
    if (x > 190 and x < 210 and y > 90 and y < 110) != (x > 190 and x < 210 and y * 4 > 90 and y < 110): b[1] = 2
    if (x > 190 and x < 210 and y > 90 and y < 110) != (x > 190 and x < 210 and y * 8 > 90 and y < 110): b[2] = 3
    if (y > 85 and y < 115 and z > 1.25 and z < 1.45) != (y > 85 and y < 115 and z > 1.25 and z < 15): b[3] = 4
    if (x > 195 and x < 205) != (x * 8 > 195 and x < 205): b[4] = 5
    if (x > 195 and x < 205) != (x > 195 and x * 8 < 205): b[5] = 6
    if (x > 195 and x < 205) != (x > 195 and x < 605): b[6] = 7
    if (y > 95 and y < 105) != (y > 95 and y * 8 < 105): b[7] = 8
    if (z > 1.3 and z < 1.4) != (z > 1.3 and z < 14): b[8] = 9
    if (z > 1 and z < 3.4) != (z > 1 and z < 4.4): b[9] = 10

    # 规则11-20
    if (abs(x - 200) < 8) != (abs(x - 200) < 18): b[10] = 11
    if (abs(y - 100) < 8) != (abs(y ** 2 - 100) < 8): b[11] = 12
    if (abs(z - 1.35) < 0.08) != (abs(z - 1.35) < 1.08): b[12] = 13
    if (abs(z - 1.35) < 0.08) != (abs(z - 1.35) < 2.08): b[13] = 14
    if (abs(z - 1.35) < 1.08) != (abs(z + 1.35) < 1.08): b[14] = 15
    if (x > 188 and x < 212 and y > 88 and y < 112 and z > 1.28 and z < 1.42) != (
            x > 188 and x < 212 and y > 88 and y < 112 and z > 1.28 and z < 9.42): b[15] = 16
    if ((x + y + z * 100) > 420 and (x + y + z * 100) < 460) != ((x + y + z * 100) > 720 and (x + y + z * 100) < 460):
        b[16] = 17
    if ((x + y + z * 100) > 420 and (x + y + z * 100) < 460) != ((x + y + z * 100) > 420 and (x + y + z * 100) < 260):
        b[17] = 18
    if ((x + y + z * 100) > 420 and (x + y + z * 100) < 460) != ((x + y + z * 100) > 420 and (x + y + z * 100) < 360):
        b[18] = 19
    if ((x + y + z * 100) > 420 and (x + y + z * 100) < 460) != ((x + y + z * 100) > 420 and (x + y + z * 100) < 560):
        b[19] = 20

    # 规则21-30
    if (x / y > 1.8 and x / y < 2.2) != (x / y > 1.8 and x / y < 22): b[20] = 21
    if (x / y > 1.8 and x / y < 2.2) != (x / y > 1.8 and x / y < 12.2): b[21] = 22
    if (x / z > 130 and x / z < 170) != (x / z > 130 and x / z < 1170): b[22] = 23
    if (x / z > 130 and x / z < 170) != (x / z > 130 and x / z < 120): b[23] = 24
    if (x / z > 130 and x / z < 170) != (x / z > 130 and x / z < 270): b[24] = 25
    if (y / z > 60 and y / z < 90) != (y / z > 60 and y / z < 910): b[25] = 26
    if (y / z > 60 and y / z < 90) != (y / z > 60 and y / z < 190): b[26] = 27
    if (y / z > 60 and y / z < 90) != (y / z > 60 and y / z < 60): b[27] = 28
    if (x * y > 16000 and x * y < 24000) != (x * y > 16000 and x * y < 2400): b[28] = 29
    if (x * y > 16000 and x * y < 24000) != (x * y > 16000 and x * y < 12400): b[29] = 30

    # 规则31-40
    if (x * y > 16000 and x * y < 24000) != (x * y > 16000 and 22 * y < 24000): b[30] = 31
    if (x * z > 240 and x * z < 320) != (x * z > 240 and x * 22 < 320): b[31] = 32
    if (x * z > 240 and x * z < 320) != (x * z > 240 and 22 * z < 320): b[32] = 33
    if (x * z > 240 and x * z < 320) != (x * z > 240 and x * 99 < 320): b[33] = 34
    if (y * z > 100 and y * z < 170) != (y * z > 100 and y * 22 < 170): b[34] = 35
    if (y * z > 100 and y * z < 170) != (y * z > 100 and 122 * z < 170): b[35] = 36
    if (y * z > 100 and y * z < 170) != (y * z > 100 and y * 88 < 170): b[36] = 37
    if ((x + y) / 2 > 145 and (x + y) / 2 < 155) != ((x + y) / 2 > 145 and (x + y) / 4 < 155): b[37] = 38
    if ((x + y) / 2 > 145 and (x + y) / 2 < 155) != ((x + y) / 2 > 145 and (x + y) / 2 < 455): b[38] = 39
    if ((x + y) / 2 > 145 and (x + y) / 2 < 155) != ((x + y) / 2 > 145 and (x + y) / 2 < 855): b[39] = 40

    # 规则41-50
    if (x - y > 60 and x - y < 120) != (x - y > 60 and x - y < 420): b[40] = 41
    if (x - y > 60 and x - y < 120) != (x - y > 60 and x - y < 620): b[41] = 42
    if (x - y > 60 and x - y < 120) != (x - y > 60 and x - y < 820): b[42] = 43
    if (abs(x / y - 2.0) < 0.3) != (abs(x / y - 2.0) < 1.3): b[43] = 44
    if (abs(x / y - 2.0) < 0.3) != (abs(x / y - 2.0) < 2.3): b[44] = 45
    if (abs(x / y - 2.0) < 0.3) != (abs(x / y - 2.0) < 55.3): b[45] = 46
    if (x / (y + 50) > 1.2 and x / (y + 50) < 1.8) != (x / (y + 50) > 1.2 and x / (y + 50) < 18): b[46] = 47
    if (x / (y + 50) > 1.2 and x / (y + 50) < 1.8) != (x / (y + 50) > 1.2 and x / (y + 50) < 221.8): b[47] = 48
    if (x / (y + 50) > 1.2 and x / (y + 50) < 1.8) != (x / (y + 50) > 1.2 and x / (y + 50) < 118): b[48] = 49
    if (y / (z * 50) > 1.2 and y / (z * 50) < 1.8) != (y / (z * 50) > 1.2 and y / (z * 50) < 1118): b[49] = 50

    # 规则51-60
    if (y / (z * 50) > 1.2 and y / (z * 50) < 1.8) != (y / (z * 50) > 1.2 and y / (z * 50) < 221.8): b[50] = 51
    if (y / (z * 50) > 1.2 and y / (z * 50) < 1.8) != (y / (z * 50) > 1.2 and y / (z * 50) < 122.8): b[51] = 52
    if (z / (x / 200) > 1.2 and z / (x / 200) < 1.8) != (z / (x / 200) > 1.2 and z / (x / 200) < 111.8): b[52] = 53
    if (z / (x / 200) > 1.2 and z / (x / 200) < 1.8) != (z / (x / 200) > 1.2 and z / (x / 200) < 18): b[53] = 54
    if (z / (x / 200) > 1.2 and z / (x / 200) < 1.8) != (z / (x / 200) > 1.2 and z / (x / 200) < 1.811): b[54] = 55
    if ((x * y * z) > 25000 and (x * y * z) < 35000) != ((x * y * z) > 25000 and (x * y * z) < 300): b[55] = 56
    if ((x + y + z) / 3 > 108 and (x + y + z) / 3 < 118) != ((x + y + z) / 3 > 108 and (x + y + z) / 5 < 118): b[56] = 57
    if ((x + y + z) / 3 > 108 and (x + y + z) / 3 < 118) != ((x + y + z) / 3 > 108 and (x + y + z) / 9 < 118): b[57] = 58
    if ((x + y + z) / 3 > 108 and (x + y + z) / 3 < 118) != ((x + y + z) / 3 > 108 and (x + y + z) / 8 < 118): b[58] = 59
    if (x ** 2 + y ** 2 + z ** 2 * 10000 > 50000) != (x ** 2 + y ** 2 + z ** 2 * 10000 > 80000): b[59] = 60

    # 规则61-70
    if (x ** 2 + y ** 2 + z ** 2 * 10000 > 50000) != (x ** 2 + y ** 2 + z ** 2 * 10000 > 99000): b[60] = 61
    if (x ** 2 + y ** 2 + z ** 2 * 10000 > 50000) != (x ** 2 + y ** 2 + z ** 2 * 10000 > 88000): b[61] = 62
    if ((x - 180) * (y - 80) > 200 and (x - 180) * (y - 80) < 1000) != (
            (x - 180) * (y - 80) > 200 and (x - 180) * (y - 80) < 4000): b[62] = 63
    if (abs((x + y) - 300) < 20) != (abs((x + y) - 300) < 120): b[63] = 64
    if (abs((x + y) - 300) < 20) != (abs((x + y) - 300) < 210): b[64] = 65
    if (abs((x + y) - 300) < 20) != (abs((x + y) - 300) < 2990): b[65] = 66
    if (x / 200 > 0.9 and x / 200 < 1.1) != (x / 200 > 0.9 and x / 200 < 11): b[66] = 67
    if (x / 200 > 0.9 and x / 200 < 1.1) != (x / 200 > 0.9 and x / 200 < 8.1): b[67] = 68
    if (x / 200 > 0.9 and x / 200 < 1.1) != (x / 200 > 0.9 and x / 200 < 99.1): b[68] = 69
    if (y / 100 > 0.9 and y / 100 < 1.1) != (y / 100 > 0.9 and y / 100 < 11): b[69] = 70

    # 规则71-80
    if (y / 100 > 0.9 and y / 100 < 1.1) != (y / 100 > 0.9 and y / 10 < 1.1): b[70] = 71
    if (y / 100 > 0.9 and y / 100 < 1.1) != (y / 100 > 0.9 and y / 200 < 1.1): b[71] = 72
    if (x < 175) != (x < 115): b[72] = 73
    if (x < 175) != (x < 125): b[73] = 74
    if (x > 225) != (x > 2215): b[74] = 75
    if (x > 225) != (x > 2299): b[75] = 76
    if (x > 225) != (x > 1225): b[76] = 77
    if (y < 75) != (y < 175): b[77] = 78
    if (y < 75) != (y < 715): b[78] = 79
    if (y < 75) != (y < 751): b[79] = 80

    # 规则81-90
    if (y > 125) != (y > 925): b[80] = 81
    if (y > 125) != (y > 1115): b[81] = 82
    if (y > 125) != (y > 1215): b[82] = 83
    if (z < 1.15) != (z < 3.85): b[83] = 84
    if (z < 1.15) != (z < 4.5): b[84] = 85
    if (z < 2.15) != (z < 0.15): b[85] = 86
    if (z > 2.55) != (z > 1.25): b[86] = 87
    if (z > 1.55) != (z > 7.45): b[87] = 88
    if (z > 1.55) != (z > 165): b[88] = 89
    if (x < 170 or x > 230) != (x < 170 or x > 1230): b[89] = 90

    # 规则91-100
    if (x < 170 or x > 230) != (x < 170 or x > 2130): b[90] = 91
    if (x < 165 and y < 75) != (x < 165 and y * 8 < 75): b[91] = 92
    if (x > 235 and y > 125) != (x > 235 and y ** 9 > 125): b[92] = 93
    if (x > 235 and y > 125) != (x > 235 and y ** 2 > 125): b[93] = 94
    if (x > 235 and y > 125) != (x > 235 and y ** 4 > 125): b[94] = 95
    if (x < 170 and z < 1.15) != (x < 170 and z < 4.15): b[95] = 96
    if (x < 170 and z < 1.15) != (x < 170 and z < 9.15): b[96] = 97
    if (x < 170 and z < 1.15) != (x < 170 and z < 94.15): b[97] = 98
    if (x > 230 and z > 1.55) != (x > 230 and z > 155): b[98] = 99
    if (x > 230 and z > 1.55) != (x > 230 and z > 4.55): b[99] = 100

    # 规则101-110
    if (y > 125 and z > 1.55) != (y > 125 and z * 2 > 1.55): b[100] = 101
    if (y > 125 and z > 1.55) != (y > 125 and z * 6 > 1.55): b[101] = 102
    if (abs(x / y - 2.0) > 0.5) != (abs(x / y - 2.0) > 10.5): b[102] = 103
    if (abs(x / y - 2.0) > 0.5) != (abs(x / y - 2.0) > 120.5): b[103] = 104
    if (abs(x / y - 2.0) > 0.5) != (abs(x / y - 2.0) > 5): b[104] = 105
    if (x * y < 14000 or x * y > 26000) != (x * y < 14000 or x * y > 28000): b[105] = 106
    if (x * y < 14000 or x * y > 26000) != (x * y < 14000 or x * y > 27000): b[106] = 107
    if (x * y < 14000 or x * y > 26000) != (x * y < 14000 or x * y > 30000): b[107] = 108
    if ((x + y + z * 100) < 400 or (x + y + z * 100) > 480) != ((x + y + z * 100) < 400 or (x + y + z * 100) > 4180): b[108] = 109
    if ((x + y + z * 100) < 400 or (x + y + z * 100) > 480) != ((x + y + z * 100) < 400 or (x + y + z * 1100) > 480): b[109] = 110

    # 规则111-120
    if ((x + y + z * 100) < 400 or (x + y + z * 100) > 480) != ((x + y + z * 100) < 400 or (x + y + z * 800) > 480): b[110] = 111
    if (x < 160 or x > 240) != (x < 160 or x > 1240): b[111] = 112
    if (x < 160 or x > 240) != (x < 160 or x > 2140): b[112] = 113
    if (x < 160 or x > 240) != (x < 160 or x > 2410): b[113] = 114
    if (y < 60 or y > 140) != (y < 60 or y > 1410): b[114] = 115
    if (y < 60 or y > 140) != (y < 60 or y > 1420): b[115] = 116
    if (y < 60 or y > 140) != (y < 60 or y > 1770): b[116] = 117
    if (z > 1.33 and z < 1.37) != (z > 1.33 and z < 137): b[117] = 118
    if (abs(x - 200) < 3 and abs(y - 100) < 3) != (abs(x - 200) < 3 and abs(y - 100) < 13): b[118] = 119
    if (abs(x - 200) < 3 and abs(y - 100) < 3) != (abs(x - 200) < 3 and abs(y - 100) < 322): b[119] = 120

    # 规则121-130
    if (abs(x - 200) < 3 and abs(y - 100) < 3) != (abs(x - 200) < 3 and abs(y - 100) < 31): b[120] = 121
    if (abs(z - 1.35) < 0.03) != (abs(z - 1.35) < 4.03): b[121] = 122
    if (abs(z - 1.35) < 0.03) != (abs(z - 1.35) < 8.03): b[122] = 123
    if (abs(z - 1.35) < 2.03) != (abs(z - 1.35) < 1.3): b[123] = 124
    if (x / y > 1.95 and x / y < 2.05) != (x / y > 125 and x / y < 2.05): b[124] = 125
    if (x / y > 1.95 and x / y < 2.05) != (x / y > 1.95 and x / y < 205): b[125] = 126
    if (x / y > 1.95 and x / y < 2.05) != (x / y > 1.95 and x / y < 2105): b[126] = 127
    if (x * y > 19000 and x * y < 21000) != (x * y > 19000 and x * y < 1000): b[127] = 128
    if (x * y > 19000 and x * y < 21000) != (x * y > 1900 and x * y < 21000): b[128] = 129
    if (x * y > 19000 and x * y < 21000) != (x * y > 19000 and x * y < 21090): b[129] = 130

    # 规则131-140
    if ((x + y + z * 100) > 430 and (x + y + z * 100) < 440) != ((x + y + z * 100) > 430 and (x + y + z * 100) < 4140): b[130] = 131
    if ((x + y + z * 100) > 430 and (x + y + z * 100) < 440) != ((x + y + z * 100) > 430 and (x + y + z * 100) < 1440): b[131] = 132
    if ((x + y + z * 100) > 430 and (x + y + z * 100) < 440) != ((x + y + z * 100) > 430 and (x + y + z * 100) < 2440): b[132] = 133
    if ((x * y * z) > 29000 and (x * y * z) < 31000) != ((x * y * z) > 29000 and (x * y * z) < 41000): b[133] = 134
    if ((x * y * z) > 29000 and (x * y * z) < 31000) != ((x * y * z) > 29000 and (x * y * z) < 38000): b[134] = 135
    if ((x * y * z) > 29000 and (x * y * z) < 31000) != ((x * y * z) > 29000 and (x * y * z) < 39000): b[135] = 136
    if (((x - 200) ** 2 + (y - 100) ** 2) < 25) != (((x - 200) ** 2 + (y - 100) ** 2) < 125): b[136] = 137
    if (((x - 200) ** 2 + (y - 100) ** 2) < 25) != (((x - 200) ** 2 + (y - 100) ** 2) < 215): b[137] = 138
    if (((x - 200) ** 2 + (y - 100) ** 2) < 25) != (((x - 200) ** 2 + (y - 100) ** 2) < 251): b[138] = 139
    if (abs((x + y) / 2 - 150) < 2) != (abs((x + y) / 2 - 150) < 12): b[139] = 140

    # 规则141-150
    if (abs((x + y) / 2 - 150) < 2) != (abs((x + y) / 2 - 150) < 21): b[140] = 141
    if (abs((x + y) / 2 - 150) < 2) != (abs((x + y) / 2 - 150) < 9): b[141] = 142
    if (x / (z * 100) > 1.45 and x / (z * 100) < 1.55) != (x / (z * 100) > 1.45 and x / (z * 100) < 11.55): b[142] = 143
    if (y / (z * 50) > 1.45 and y / (z * 50) < 1.55) != (y / (z * 50) > 1.45 and y / (z * 50) < 55): b[143] = 144
    if ((x - 180) / 40 > 0.45 and (x - 180) / 40 < 0.55) != ((x - 180) / 40 > 0.45 and (x - 180) / 4 < 0.55): b[144] = 145
    if ((x - 180) / 40 > 0.45 and (x - 180) / 40 < 0.55) != ((x - 180) / 40 > 0.45 and (x - 180) / 90 < 0.55): b[145] = 146
    if ((x - 180) / 40 > 0.45 and (x - 180) / 40 < 0.55) != ((x - 180) / 40 > 0.45 and (x - 180) / 80 < 0.55): b[146] = 147
    if ((y - 80) / 40 > 0.45 and (y - 80) / 40 < 0.55) != ((y - 80) / 40 > 0.45 and (y - 80) / 40 < 22.55): b[147] = 148
    if ((y - 80) / 40 > 0.45 and (y - 80) / 40 < 0.55) != ((y - 80) / 40 > 0.45 and (y - 80) / 40 < 2.55): b[148] = 149
    if ((y - 80) / 40 > 0.45 and (y - 80) / 40 < 0.55) != ((y - 80) / 40 > 0.45 and (y - 80) / 40 < 6.55): b[149] = 150

    # 规则151-160
    if ((z - 1.2) / 0.3 > 0.45 and (z - 1.2) / 0.3 < 0.55) != ((z - 1.2) / 0.3 > 0.45 and (z - 1.2) / 3 < 0.55): b[150] = 151
    if ((z - 1.2) / 0.3 > 0.45 and (z - 1.2) / 0.3 < 0.55) != ((z - 1.2) / 0.3 > 0.45 and (z - 1.2) / 0.3 < 5.55): b[151] = 152
    if ((z - 1.2) / 0.3 > 0.45 and (z - 1.2) / 0.3 < 0.55) != ((z - 1.2) / 0.3 > 0.45 and (z - 1.2) / 3 < 0.55): b[152] = 153
    if (min(x - 180, 220 - x) > 15) != (min(x - 180, 220 - x) > 152): b[153] = 154
    if (min(x - 180, 220 - x) > 15) != (min(x - 180, 220 - x) > 115): b[154] = 155
    if (min(x - 180, 220 - x) > 15) != (min(x - 180, 220 - x) > 5): b[155] = 156
    if (min(y - 80, 120 - y) > 15) != (min(y - 80, 120 + y) > 15): b[156] = 157
    if (min(y - 80, 120 - y) > 15) != (min(y - 80, 120 * y) > 15): b[157] = 158
    if (min(y - 80, 120 - y) > 15) != (min(y - 80, 120 - y) > 95): b[158] = 159
    if (min(y - 80, 120 - y) > 15) != (min(y - 80, 120 - y) > 35): b[159] = 160

    # 规则161-163
    if (abs(x * y - 20000) < 500) != (abs(x * y - 20000) < 1500): b[160] = 161
    if (abs(x * y - 20000) < 500) != (abs(x * y - 20000) < 5100): b[161] = 162
    if (abs(x * y - 20000) < 500) != (abs(x * y - 20000) < 5010): b[162] = 163

    # 将触发的规则添加到结果集合中
    for key, value in b.items():
        triggered.add(value)

    return triggered


def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """"""
    generated_path = execute_validation_rules(particle[0], particle[1], particle[2])

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # : x:1-50, y:1-50, z:1-50
        self.bounds = bounds if bounds else [(1, 300), (1, 200), (1, 5)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """target pathsPSO"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(particles[i][0], particles[i][1], particles[i][2]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(gbest_particle[0], gbest_particle[1], gbest_particle[2]),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """Path PSO"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - Path ")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations")
    print(f"Path : {len(target_paths)}")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"Path {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "" if result['success'] else f"({result['best_fitness']:.3f})"
        print(f"{status} | {result['time']:.2f}s | iterations{result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f": {success_count}/{len(target_paths)} ({success_rate:.1f}%) | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs}")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations, {len(target_paths)}Path ")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"---  {run_idx}/{num_runs} ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f": {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs} runcompleted | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 1: 
    ws1 = wb.active
    ws1.title = ""
    ws1.sheet_view.showGridLines = False

    headers = ["Run", "", "", "", "Average Iterations", "(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f" {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # 2: Path 
    ws2 = wb.create_sheet(title="Path ")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Path ID", "", "", "", "Average Iterations", "Minimum Iterations", "Maximum Iterations"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"Path  {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # 3: 
    ws3 = wb.create_sheet(title="")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Path ", "", "(x,y,z)", "", "Iterations", "Path "]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"{run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # 4: target paths
    ws4 = wb.create_sheet(title="target paths")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Path ID", "target paths", ""]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"Path  {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f" : {filename}")
    print(f"{'=' * 70}")
    print(f":")
    print(f"  1.        - {len(all_results)} run")
    print(f"  2. Path        - Path ")
    print(f"  3.    -  runPath ")
    print(f"  4. target paths       - target paths")
    print(f"{'=' * 70}\n")

    return filename


def main():


    target_paths = [
        {7, 8, 12, 13, 14, 15, 20, 21, 22, 23, 25, 26, 27, 31, 32, 34, 35, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 50,
         51,
         52, 56, 57, 58, 59, 61, 63, 64, 65, 66, 67, 68, 69, 71, 75, 76, 77, 78, 79, 80, 86, 90, 91, 93, 94, 95, 103,
         104,
         105, 106, 107, 108, 110, 111, 112, 113, 114, 122, 123, 126, 127, 131, 132, 133, 143, 144, 148, 149, 150, 159,
         160},

        {7, 8, 12, 13, 14, 15, 17, 18, 19, 21, 22, 23, 25, 26, 27, 31, 32, 34, 35, 37, 38, 39, 40, 41, 42, 43, 44, 45,
         46,
         50, 51, 52, 56, 57, 58, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 75, 76, 77, 78, 79, 80, 86, 90, 91, 93, 94,
         95,
         103, 104, 105, 106, 107, 108, 110, 111, 112, 113, 114, 122, 123, 126, 127, 131, 132, 133, 143, 144, 159, 160},

        {7, 13, 14, 15, 20, 21, 22, 23, 26, 27, 31, 32, 34, 35, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50,
         51,
         52, 56, 57, 58, 59, 63, 64, 65, 66, 67, 68, 69, 70, 72, 75, 76, 77, 78, 79, 80, 86, 90, 91, 93, 94, 95, 103,
         104,
         105, 109, 112, 113, 114, 122, 123, 126, 127, 131, 132, 133, 134, 135, 136, 143, 144, 148, 149, 150, 157, 158},

        {4, 7, 8, 9, 12, 13, 14, 15, 21, 22, 24, 31, 33, 38, 39, 40, 41, 42, 43, 44, 45, 46, 57, 58, 59, 63, 64, 65, 66,
         67,
         68, 69, 71, 75, 76, 77, 78, 79, 80, 84, 85, 86, 87, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 106,
         107,
         108, 109, 112, 113, 114, 118, 122, 123, 126, 127, 131, 132, 133, 148, 149, 150, 151, 152, 153, 159, 160},

        {7, 9, 13, 14, 15, 20, 21, 22, 29, 30, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 64, 65,
         66,
         67, 68, 69, 75, 76, 77, 78, 79, 80, 84, 85, 86, 87, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 109,
         112,
         113, 114, 118, 122, 123, 126, 127, 128, 131, 132, 133, 134, 136, 141, 151, 152, 153, 161, 162, 163},

        {7, 9, 13, 14, 15, 20, 21, 22, 24, 29, 30, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 47, 48, 49, 57, 58, 59,
         64,
         65, 66, 67, 68, 69, 75, 76, 77, 84, 85, 86, 87, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 109, 112,
         113,
         114, 118, 122, 123, 126, 127, 129, 131, 132, 133, 134, 135, 136, 151, 152, 153, 161, 162, 163},

        {7, 9, 13, 14, 15, 20, 21, 22, 24, 29, 30, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 64,
         65,
         66, 67, 68, 69, 75, 76, 77, 78, 79, 80, 84, 85, 86, 87, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105,
         109,
         112, 113, 114, 118, 122, 123, 126, 127, 130, 131, 132, 133, 151, 152, 153, 161, 162, 163},

        {7, 9, 13, 14, 15, 20, 21, 22, 29, 30, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 67, 68,
         69,
         75, 76, 77, 78, 79, 80, 84, 85, 86, 87, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 109, 118, 122, 123,
         126, 127, 129, 131, 132, 133, 134, 135, 136, 140, 141, 142, 151, 152, 153, 161, 162, 163},

        {7, 13, 14, 15, 20, 23, 25, 26, 27, 31, 32, 34, 35, 37, 38, 39, 40, 44, 45, 46, 50, 51, 52, 56, 57, 58, 59, 61,
         63,
         64, 65, 66, 67, 68, 69, 70, 72, 75, 76, 77, 78, 79, 80, 81, 82, 83, 86, 90, 91, 101, 102, 109, 112, 113, 114,
         115,
         116, 117, 122, 123, 131, 132, 133, 134, 135, 136, 143, 144, 148, 149, 150, 157, 158},

        {7, 11, 13, 14, 15, 17, 18, 19, 23, 25, 26, 27, 31, 35, 37, 38, 39, 40, 44, 45, 46, 50, 51, 52, 56, 57, 58, 59,
         60,
         61, 62, 63, 64, 65, 66, 70, 72, 78, 79, 80, 81, 82, 83, 86, 101, 102, 103, 104, 105, 110, 111, 115, 116, 117,
         122,
         123, 131, 132, 133, 134, 135, 136, 143, 144, 146, 147, 148, 149, 150, 156, 157, 158},

        {5, 9, 13, 14, 15, 20, 28, 29, 30, 32, 34, 44, 45, 46, 53, 54, 60, 61, 62, 64, 65, 66, 70, 72, 73, 78, 79, 80,
         81,
         82, 83, 84, 85, 86, 87, 88, 89, 96, 97, 98, 103, 104, 105, 110, 111, 115, 116, 117, 118, 122, 123, 128, 131,
         132,
         133, 134, 136, 140, 141, 144, 148, 149, 150, 151, 152, 153, 157, 158, 161, 162, 163},

        {5, 9, 13, 14, 15, 20, 28, 29, 30, 32, 34, 44, 45, 46, 53, 54, 60, 61, 62, 64, 65, 66, 70, 72, 73, 74, 78, 79,
         80,
         81, 82, 83, 84, 85, 86, 87, 88, 89, 96, 97, 98, 103, 104, 105, 110, 111, 115, 116, 117, 118, 122, 123, 128,
         131,
         132, 133, 134, 136, 140, 141, 148, 149, 150, 151, 152, 153, 157, 158, 161, 162, 163},

        {4, 7, 8, 9, 12, 14, 21, 22, 31, 33, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 53, 54, 57, 58, 59, 63, 64,
         65,
         66, 67, 68, 69, 71, 75, 76, 77, 78, 79, 80, 84, 85, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 106,
         107,
         108, 109, 112, 113, 114, 118, 122, 123, 124, 126, 127, 131, 132, 133, 159, 160},

        {6, 13, 14, 15, 17, 18, 19, 23, 25, 26, 27, 31, 35, 37, 38, 39, 40, 44, 45, 46, 50, 51, 52, 56, 57, 58, 59, 60,
         61,
         62, 63, 64, 65, 66, 70, 72, 78, 79, 80, 81, 82, 83, 86, 101, 102, 103, 104, 105, 110, 111, 115, 116, 117, 120,
         122,
         123, 131, 132, 133, 143, 144, 146, 147, 148, 149, 150, 154, 155, 157, 158},

        {6, 13, 14, 15, 17, 18, 19, 23, 25, 26, 27, 31, 35, 37, 38, 39, 40, 44, 45, 46, 50, 51, 52, 56, 60, 61, 62, 63,
         64,
         65, 66, 70, 72, 78, 79, 80, 81, 82, 83, 86, 101, 102, 103, 104, 105, 108, 110, 111, 115, 116, 117, 120, 122,
         123,
         131, 132, 133, 141, 143, 144, 145, 148, 149, 150, 154, 155, 157, 158},

        {4, 7, 9, 10, 12, 21, 22, 31, 33, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 53, 54, 57, 58, 59, 63, 64,
         65,
         66, 67, 68, 69, 71, 75, 76, 77, 78, 79, 80, 85, 88, 89, 90, 91, 93, 94, 95, 99, 100, 103, 104, 105, 109, 112,
         113,
         114, 118, 122, 123, 126, 127, 131, 132, 133, 148, 149, 150, 157, 158},

        {7, 13, 14, 15, 20, 23, 26, 27, 31, 32, 34, 35, 37, 38, 39, 40, 41, 42, 43, 50, 51, 52, 57, 58, 59, 65, 66, 67,
         68,
         69, 70, 72, 75, 76, 77, 78, 79, 80, 81, 82, 83, 86, 90, 91, 101, 102, 109, 112, 113, 114, 115, 116, 117, 122,
         123,
         125, 131, 132, 133, 134, 143, 144, 148, 149, 150, 157, 158},

        {7, 9, 13, 14, 15, 28, 31, 33, 38, 39, 40, 44, 45, 46, 53, 54, 55, 57, 58, 59, 63, 64, 65, 66, 67, 68, 69, 70,
         72,
         78, 79, 80, 81, 82, 83, 84, 85, 86, 87, 88, 89, 103, 104, 105, 109, 115, 116, 117, 118, 122, 123, 131, 132,
         133,
         144, 146, 147, 148, 149, 150, 151, 152, 153, 157, 158},

        {1, 2, 3, 7, 9, 13, 14, 15, 20, 21, 22, 29, 30, 33, 35, 36, 37, 41, 42, 43, 44, 45, 46, 53, 54, 56, 61, 78, 79,
         80,
         84, 85, 86, 87, 88, 89, 103, 104, 105, 109, 118, 122, 123, 126, 127, 129, 131, 132, 133, 134, 135, 136, 140,
         141,
         142, 146, 147, 151, 152, 153, 156, 162, 163},

        {1, 2, 3, 4, 6, 9, 13, 14, 15, 16, 20, 21, 22, 29, 30, 33, 53, 54, 61, 78, 79, 80, 84, 85, 86, 87, 88, 89, 109,
         118,
         119, 120, 121, 122, 123, 126, 127, 129, 131, 132, 133, 134, 135, 136, 137, 138, 139, 140, 141, 142, 145, 151,
         152,
         153, 154, 155, 162, 163},

        {5, 9, 13, 14, 15, 17, 18, 19, 21, 22, 32, 34, 35, 36, 37, 44, 45, 46, 53, 54, 60, 61, 62, 64, 65, 66, 73, 74,
         84,
         85, 86, 87, 88, 89, 92, 96, 97, 98, 110, 111, 118, 122, 123, 126, 127, 129, 151, 152, 153},
    ]

    print("=" * 70)
    print("baseline PSO")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("Program completed")


if __name__ == "__main__":
    main()