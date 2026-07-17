import os
import random
import math
import numpy as np
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# --- 状态范围配置（按 category5 注释）---
STATE_MIN_X, STATE_MAX_X = 1, 250   # 温度/光照
STATE_MIN_Y, STATE_MAX_Y = 1, 200   # 电压/湿度
STATE_MIN_Z, STATE_MAX_Z = 1, 5     # 流量/电流因子

def generate_input():
    return [
        random.randint(STATE_MIN_X, STATE_MAX_X),
        random.randint(STATE_MIN_Y, STATE_MAX_Y),
        random.randint(STATE_MIN_Z, STATE_MAX_Z)
    ]

# ========== 规则函数 category5（修正版） ==========
def category5_multivariable_control(x, y, z):
    # 固定值设置: x:1--250, y:1--200, z:1--5
    triggered = set()

    # 分支1-10: 光照与温度协同控制
    if (x > 190 and x < 210 and y > 45 and y < 100) != (x > 150 and x < 210 and y > 45 and y < 100): triggered.add(1)
    if (x > 190 and x < 210 and y > 45 and y < 100) != (x > 100 and x < 210 and y > 45 and y < 100): triggered.add(2)
    if (x > 190 and x < 210 and y > 45 and y < 100) != (x > 120 and x < 210 and y > 45 and y < 100): triggered.add(3)

    if (x > 105 and x < 215 and z > 1 and z < 3) != (x > 125 and x < 215 and z > 1 and z < 3): triggered.add(4)
    if (x > 105 and x < 215 and z > 1 and z < 3) != (x > 155 and x < 215 and z > 1 and z < 3): triggered.add(5)
    if (x > 105 and x < 215 and z > 1 and z < 3) != (x > 105 and x < 245 and z > 1 and z < 3): triggered.add(6)
    if (x > 105 and x < 215 and z > 1 and z < 3) != (x > 105 and x < 215 and z > 1 and z < 4): triggered.add(7)
    if (x > 105 and x < 215 and z > 1 and z < 3) != (y > 105 and x < 215 and z > 1 and z < 3): triggered.add(8)
    if (x > 105 and x < 215 and z > 1 and z < 3) != (x > 105 or x < 215 and z > 1 and z < 3): triggered.add(9)

    if (y > 40 and y < 150 and z > 1 and z < 4) != (x > 40 and y < 150 and z > 1 and z < 4): triggered.add(10)
    if (y > 40 and y < 150 and z > 1 and z < 4) != (y > 40 and x < 150 and z > 1 and z < 4): triggered.add(11)
    if (y > 40 and y < 150 and z > 1 and z < 4) != (y > 60 and y < 150 and z > 1 and z < 4): triggered.add(12)
    if (y > 40 and y < 150 and z > 1 and z < 4) != (y > 80 and y < 150 and z > 1 and z < 4): triggered.add(13)
    if (y > 40 and y < 150 and z > 1 and z < 4) != (y > 40 and y < 170 and z > 1 and z < 4): triggered.add(14)
    if (y > 40 and y < 150 and z > 1 and z < 4) != (y > 40 and y < 150 and z > 1 or z < 4): triggered.add(15)
    if (y > 40 and y < 150 and z > 1 and z < 4) != (y > 40 and y < 150 and z > 1 and z > 4): triggered.add(16)
    if (y > 40 and y < 150 and z > 1 and z < 4) != (y > 40 and y < 150 and z > 2 and z < 4): triggered.add(17)
    if (y > 40 and y < 150 and z > 1 and z < 4) != (y > 40 and y < 150 or z > 1 and z < 4): triggered.add(18)

    if (x > 100 and x < 150) != (x > 120 and x < 150): triggered.add(19)
    if (x > 100 and x < 150) != (x > 140 and x < 150): triggered.add(20)
    if (x > 100 and x < 150) != (x > 50 and x < 150): triggered.add(21)
    if (x > 100 and x < 150) != (x > 100 and x < 200): triggered.add(22)
    if (x > 100 and x < 150) != (x > 100 and x < 210): triggered.add(23)
    if (x > 100 and x < 150) != (x < 100 and x < 150): triggered.add(24)
    if (x > 100 and x < 150) != (x > 100 and x > 150): triggered.add(25)
    if (x > 100 and x < 150) != (x > 100 and y < 150): triggered.add(26)

    if (y > 108 and y < 152) != (x > 108 and y < 152): triggered.add(27)
    if (y > 108 and y < 152) != (y > 108 and x < 152): triggered.add(28)
    if (y > 108 and y < 152) != (y > 78 and y < 152): triggered.add(29)
    if (y > 108 and y < 152) != (y > 48 and y < 152): triggered.add(30)
    if (y > 108 and y < 152) != (y > 108 and y < 142): triggered.add(31)
    if (y > 108 and y < 152) != (y > 108 and y < 132): triggered.add(32)

    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 100 > 500 and x + y + z * 100 < 700): triggered.add(33)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 100 > 600 and x + y + z * 100 < 700): triggered.add(34)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 100 > 450 and x + y + z * 100 < 700): triggered.add(35)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 100 > 400 and x + y + z * 100 < 800): triggered.add(36)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 100 > 400 and x + y + z * 100 < 750): triggered.add(37)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 100 > 400 and x + y + z * 100 < 850): triggered.add(38)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x - y + z * 100 > 400 and x + y + z * 100 < 700): triggered.add(39)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 110 > 400 and x + y + z * 100 < 700): triggered.add(40)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 120 > 400 and x + y + z * 100 < 700): triggered.add(41)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 130 > 400 and x + y + z * 100 < 700): triggered.add(42)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 100 > 400 and x + y + z * 110 < 700): triggered.add(43)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 100 > 400 and x + y + z * 120 < 700): triggered.add(44)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 100 > 400 and x + y + z * 130 < 700): triggered.add(45)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 100 > 400 and x - y + z * 100 < 700): triggered.add(46)

    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 50 > 85 and x * y / 100 < 115): triggered.add(47)
    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 60 > 85 and x * y / 100 < 115): triggered.add(48)
    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 80 > 85 and x * y / 100 < 115): triggered.add(49)
    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 100 > 95 and x * y / 100 < 115): triggered.add(50)
    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 100 > 105 and x * y / 100 < 115): triggered.add(51)
    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 100 > 85 or x * y / 100 < 115): triggered.add(52)
    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 100 > 85 and x * y / 100 < 125): triggered.add(53)
    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 100 > 85 and x * y / 100 < 135): triggered.add(54)
    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 100 > 85 and x * y / 100 < 155): triggered.add(55)
    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 100 > 85 and x * y / 80 < 115): triggered.add(56)
    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 100 > 85 and x * y / 70 < 115): triggered.add(57)
    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 100 > 85 and x * y / 50 < 115): triggered.add(58)
    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 100 > 85 and x * y / 60 < 115): triggered.add(59)

    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 180) + (z - 1) * 100 > 45 and (x - 180) + (z - 1) * 100 < 45): triggered.add(60)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 180) + (z - 1) * 100 > 35 and (x - 180) + (z - 1) * 100 < 45): triggered.add(61)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 180) + (z - 1) * 100 > 40 and (x - 180) + (z - 1) * 100 < 45): triggered.add(62)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 180) + (z - 1) * 100 > 25 or (x - 180) + (z - 1) * 100 < 45): triggered.add(63)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 80) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45): triggered.add(64)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 50) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45): triggered.add(65)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 180) + (z - 1) * 100 > 25 and (x - 80) + (z - 1) * 100 < 45): triggered.add(66)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 180) + (z - 1) * 100 > 25 and (x - 150) + (z - 1) * 100 < 45): triggered.add(67)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 180) + (z - 1) * 100 > 25 and (x - 180) - (z - 1) * 100 < 45): triggered.add(68)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 145): triggered.add(69)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 75): triggered.add(70)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 95): triggered.add(71)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 100): triggered.add(72)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 180) + (z - 1) * 100 > 15 and (x - 180) + (z - 1) * 100 < 45): triggered.add(73)

    if ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145) != ((y - 50) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145): triggered.add(74)
    if ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145) != ((y + 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145): triggered.add(75)
    if ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145) != ((y - 80) + (z - 2) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145): triggered.add(76)
    if ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145) != ((y - 80) + (z - 3) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145): triggered.add(77)
    if ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145) != ((y - 80) + (z - 1) * 100 > 45 and (y - 40) + (z - 1) * 50 < 145): triggered.add(78)
    if ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145) != ((y - 80) + (z - 1) * 50 > 95 and (y - 40) + (z - 1) * 50 < 145): triggered.add(79)
    if ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145) != ((y - 80) + (z - 1) * 50 > 45 and (y - 20) + (z - 1) * 50 < 145): triggered.add(80)
    if ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145) != ((y - 80) + (z - 1) * 50 > 45 and (y + 40) + (z - 1) * 50 < 145): triggered.add(81)
    if ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145) != ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 2) * 50 < 145): triggered.add(82)
    if ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145) != ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 150 < 145): triggered.add(83)
    if ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145) != ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 200): triggered.add(84)

    if (x + y * 4 + z * 160 > 540 and x + y * 4 + z * 160 < 620) != (x + y * 5 + z * 160 > 540 and x + y * 4 + z * 160 < 620): triggered.add(85)
    if (x + y * 4 + z * 160 > 540 and x + y * 4 + z * 160 < 620) != (x + y * 6 + z * 160 > 540 and x + y * 4 + z * 160 < 620): triggered.add(86)
    if (x + y * 4 + z * 160 > 540 and x + y * 4 + z * 160 < 620) != (y + y * 4 + z * 160 > 540 and x + y * 4 + z * 160 < 620): triggered.add(87)
    if (x + y * 4 + z * 160 > 540 and x + y * 4 + z * 160 < 620) != (x + y * 4 + z * 180 > 540 and x + y * 4 + z * 160 < 620): triggered.add(88)
    if (x + y * 4 + z * 160 > 540 and x + y * 4 + z * 160 < 620) != (x + y * 4 + z * 160 > 440 and x + y * 4 + z * 160 < 620): triggered.add(89)
    if (x + y * 4 + z * 160 > 540 and x + y * 4 + z * 160 < 620) != (x + y * 4 + z * 160 > 540 and x + y * 5 + z * 160 < 620): triggered.add(90)
    if (x + y * 4 + z * 160 > 540 and x + y * 4 + z * 160 < 620) != (x + y * 4 + z * 160 > 540 and x + y * 6 + z * 160 < 620): triggered.add(91)
    if (x + y * 4 + z * 160 > 540 and x + y * 4 + z * 160 < 620) != (x + y * 4 + z * 160 > 540 and x + y * 4 + z * 160 < 720): triggered.add(92)

    if (x * y > 9500 and x * y < 15500 and z > 1 and z < 4) != (x * y > 8500 and x * y < 15500 and z > 1 and z < 4): triggered.add(93)
    if (x * y > 9500 and x * y < 15500 and z > 1 and z < 4) != (x * y > 7500 and x * y < 15500 and z > 1 and z < 4): triggered.add(94)
    if (x * y > 9500 and x * y < 15500 and z > 1 and z < 4) != (x * y > 6500 and x * y < 15500 and z > 1 and z < 4): triggered.add(95)
    if (x * y > 9500 and x * y < 15500 and z > 1 and z < 4) != (x * y > 5500 and x * y < 15500 and z > 1 and z < 4): triggered.add(96)
    if (x * y > 9500 and x * y < 15500 and z > 1 and z < 4) != (x * y > 9500 and x * y < 16500 and z > 1 and z < 4): triggered.add(97)
    if (x * y > 9500 and x * y < 15500 and z > 1 and z < 4) != (x * y > 9500 and x * y < 17500 and z > 1 and z < 4): triggered.add(98)
    if (x * y > 9500 and x * y < 15500 and z > 1 and z < 4) != (x * y > 9500 and x * y < 18500 and z > 1 and z < 4): triggered.add(99)
    if (x * y > 9500 and x * y < 15500 and z > 1 and z < 4) != (x * y > 9500 and x * y < 15500 and z > 2 and z < 4): triggered.add(100)
    if (x * y > 9500 and x * y < 15500 and z > 1 and z < 4) != (x * y > 9500 and x * y < 15500 and z > 1 and z < 5): triggered.add(101)
    if (x * y > 9500 and x * y < 15500 and z > 1 and z < 4) != (x * y > 9500 and x * y < 15500 and z > 1 or z < 4): triggered.add(102)
    if (x * y > 9500 and x * y < 15500 and z > 1 and z < 4) != (x * y > 9500 and x * y < 15500 or z > 1 and z < 4): triggered.add(103)
    if (x * y > 9500 and x * y < 15500 and z > 1 and z < 4) != (x * y > 9500 or x * y < 15500 and z > 1 and z < 4): triggered.add(104)

    if ( (x + y + z * 100) < 370 or (x + y + z * 100) > 410) != ( (x + y + z * 50) < 370 or (x + y + z * 100) > 410): triggered.add(105)
    if ( (x + y + z * 100) < 370 or (x + y + z * 100) > 410) != ( (x + y + z * 60) < 370 or (x + y + z * 100) > 410): triggered.add(106)
    if ( (x + y + z * 100) < 370 or (x + y + z * 100) > 410) != ( (x + y + z * 70) < 370 or (x + y + z * 100) > 410): triggered.add(107)
    if ( (x + y + z * 100) < 370 or (x + y + z * 100) > 410) != ( (x + y + z * 80) < 370 or (x + y + z * 100) > 410): triggered.add(108)
    if ( (x + y + z * 100) < 370 or (x + y + z * 100) > 410) != ( (x + y + z * 100) < 270 or (x + y + z * 100) > 410): triggered.add(109)
    if ( (x + y + z * 100) < 370 or (x + y + z * 100) > 410) != ( (x + y + z * 100) < 370 or (x + y + z * 100) > 510): triggered.add(110)
    if ( (x + y + z * 100) < 370 or (x + y + z * 100) > 410) != ( (x + y + z * 100) < 370 or (x + y + z * 100) > 610): triggered.add(111)
    if ( (x + y + z * 100) < 370 or (x + y + z * 100) > 410) != ( (x + y + z * 100) < 370 or (x + y + z * 100) > 710): triggered.add(112)
    if ( (x + y + z * 100) < 370 or (x + y + z * 100) > 410) != ( (x + y + z * 100) < 370 or (x + y + z * 50) > 410): triggered.add(113)
    if ( (x + y + z * 100) < 370 or (x + y + z * 100) > 410) != ( (x + y + z * 100) < 370 or (x + y + z * 70) > 410): triggered.add(114)
    if ( (x + y + z * 100) < 370 or (x + y + z * 100) > 410) != ( (x + x + z * 100) < 370 or (x + y + z * 100) > 410): triggered.add(115)
    if ( (x + y + z * 100) < 370 or (x + y + z * 100) > 410) != ( (x + y + z * 100) < 370 or (x + x + z * 100) > 410): triggered.add(116)

    return triggered

# ========== 关键修正：别名指向 category5 ==========
execute_Tr = category5_multivariable_control

# ========== Jaccard 相似度 ==========
def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0

# === 目标路径组（编号与 category5 规则对应） ===
targetPaths = [
    {1, 2, 3, 5, 8, 11, 12, 13, 16, 17, 22, 23, 25, 26, 27, 30, 33, 34, 35, 39, 50, 51, 57, 58, 59, 63, 68, 69, 70, 71,
     72, 74, 75, 78, 92, 93, 94, 95, 96, 102, 103, 104, 110, 111, 112, 113, 114},

    {1, 2, 3, 8, 11, 12, 13, 16, 17, 22, 23, 25, 26, 27, 30, 33, 34, 35, 39, 50, 51, 57, 58, 59, 63, 68, 69, 71, 72, 74,
     75, 78, 92, 93, 94, 95, 96, 102, 103, 104, 110, 111, 112, 113, 114},

    {2, 4, 5, 8, 13, 16, 17, 19, 20, 24, 25, 27, 29, 30, 40, 41, 42, 50, 51, 56, 57, 58, 59, 60, 62, 66, 67, 76, 77, 79,
     81, 83, 93, 94, 95, 96, 102, 103, 104, 105, 106, 107, 108, 116},

    {2, 4, 5, 8, 16, 17, 19, 20, 24, 25, 27, 29, 30, 40, 41, 42, 50, 51, 56, 57, 58, 59, 60, 61, 62, 66, 67, 76, 77, 79,
     81, 83, 93, 94, 95, 96, 102, 103, 104, 105, 106, 107, 108, 116},

    {2, 3, 5, 8, 11, 12, 13, 16, 17, 22, 23, 26, 27, 30, 33, 34, 35, 39, 47, 48, 49, 52, 63, 68, 69, 70, 71, 72, 74, 75,
     78, 92, 94, 95, 96, 102, 103, 104, 105, 106, 107, 108, 116},

    {2, 7, 9, 13, 16, 19, 20, 24, 25, 27, 29, 30, 33, 34, 39, 50, 51, 56, 57, 58, 59, 63, 68, 69, 77, 80, 81, 83, 93,
     94, 95, 96, 102, 103, 104, 110, 111, 112, 113, 114},

    {2, 3, 4, 5, 8, 13, 16, 17, 20, 24, 25, 27, 29, 30, 33, 34, 35, 39, 51, 56, 57, 58, 59, 63, 68, 69, 70, 71, 72, 76,
     77, 79, 81, 83, 100, 105, 106, 107, 108, 116},

    {1, 2, 3, 5, 8, 11, 13, 16, 17, 22, 23, 25, 26, 27, 30, 33, 34, 35, 39, 52, 53, 54, 55, 63, 68, 69, 70, 71, 72, 76,
     77, 79, 81, 83, 100, 110, 111, 112, 113, 114},

    {1, 2, 3, 8, 11, 16, 17, 22, 23, 25, 26, 27, 29, 30, 33, 34, 39, 52, 63, 68, 69, 72, 76, 77, 79, 81, 83, 97, 98, 99,
     102, 103, 104, 110, 111, 112, 113, 114},

    {8, 9, 11, 15, 18, 21, 24, 28, 33, 34, 35, 39, 50, 51, 56, 57, 58, 59, 63, 64, 65, 82, 84, 93, 94, 95, 96, 102, 103,
     104, 110, 111, 112, 113, 114, 116},
    {9, 16, 17, 21, 24, 29, 30, 40, 41, 42, 50, 51, 56, 57, 58, 59, 63, 64, 65, 73, 76, 77, 79, 81, 83, 93, 94, 95, 96,
     102, 103, 104, 105, 106, 107, 108},
    {6, 9, 11, 12, 13, 16, 17, 25, 26, 27, 33, 34, 39, 50, 51, 57, 58, 59, 63, 68, 69, 75, 78, 92, 93, 94, 95, 96, 102,
     103, 104, 110, 111, 112, 113, 114},
    {8, 9, 16, 17, 21, 24, 27, 40, 41, 42, 50, 51, 57, 58, 59, 63, 64, 65, 76, 77, 79, 81, 83, 93, 94, 95, 96, 102, 103,
     104, 105, 106, 107, 108, 115},
    {8, 9, 16, 17, 21, 24, 27, 32, 33, 34, 35, 39, 50, 51, 56, 57, 58, 59, 63, 64, 65, 82, 84, 93, 94, 95, 96, 102, 103,
     104, 105, 106, 107, 108, 115},
    {5, 8, 10, 15, 18, 22, 23, 25, 26, 27, 40, 41, 42, 47, 48, 52, 63, 68, 69, 70, 71, 72, 75, 78, 87, 90, 91, 102, 103,
     104, 105, 106, 107, 108, 116},
    {8, 9, 11, 14, 15, 18, 21, 24, 28, 33, 34, 35, 39, 47, 48, 49, 52, 63, 64, 65, 82, 84, 94, 95, 96, 102, 103, 104,
     110, 111, 112, 113, 114, 116},
    {2, 4, 5, 8, 12, 13, 16, 17, 19, 20, 24, 25, 27, 30, 41, 42, 47, 48, 52, 60, 61, 62, 66, 67, 74, 75, 78, 92, 96,
     102, 103, 104, 109, 115},
    {4, 5, 16, 17, 19, 20, 24, 25, 31, 32, 33, 34, 39, 52, 60, 61, 62, 66, 67, 82, 84, 97, 98, 99, 102, 103, 104, 110,
     111, 112, 113, 114},
    {5, 8, 10, 15, 18, 22, 23, 25, 26, 27, 41, 42, 52, 63, 68, 69, 71, 72, 75, 85, 86, 88, 89, 102, 103, 104, 105, 106,
     107, 108, 116},
    {1, 2, 3, 9, 15, 18, 22, 23, 25, 26, 27, 29, 30, 44, 45, 52, 53, 54, 55, 63, 68, 82, 84, 101, 102, 103, 104, 112},
    {1, 2, 3, 9, 15, 16, 18, 22, 23, 25, 26, 27, 29, 30, 36, 37, 38, 46, 52, 53, 54, 55, 63, 68, 102, 103, 104},
    {2, 9, 15, 16, 18, 19, 20, 24, 25, 27, 29, 30, 43, 44, 45, 51, 56, 57, 58, 59, 63, 68, 102, 103, 104, 112},
]

# ========== 实验配置 ==========
class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 2000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(targetPaths)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }

def compute_robustness(state, path):
    base = execute_Tr(*state)
    if not base:
        return 0.0
    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                    np.clip(state[0] + dx, STATE_MIN_X, STATE_MAX_X),
                    np.clip(state[1] + dy, STATE_MIN_Y, STATE_MAX_Y),
                    np.clip(state[2] + dz, STATE_MIN_Z, STATE_MAX_Z)
                ])
                n_trig = execute_Tr(*neighbor)
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0

def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = targetPaths[target_path_idx]
    samples = []
    attempts = 0
    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(STATE_MIN_X, STATE_MAX_X),
            random.randint(STATE_MIN_Y, STATE_MAX_Y),
            random.randint(STATE_MIN_Z, STATE_MAX_Z)
        ])
        triggered = execute_Tr(*state)
        if not triggered:
            continue
        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)
        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }
        samples.append(sample_data)
    return samples

def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample['robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)
        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]
    return selected_samples

def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = targetPaths[target_path_idx]
        samples = []
        attempts = 0
        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(STATE_MIN_X, STATE_MAX_X),
                random.randint(STATE_MIN_Y, STATE_MAX_Y),
                random.randint(STATE_MIN_Z, STATE_MAX_Z)
            ])
            triggered = execute_Tr(*state)
            if not triggered:
                continue
            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)
            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)
        random.shuffle(samples)
        return samples[:config.top_k_samples]
    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")
        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)

def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}
    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates
    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}
        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config, shared_candidates[path_idx])
            strategy_results[path_idx] = samples
        results[strategy_name] = strategy_results
    return results

def analyze_fitness_values(results, config):
    analysis_results = {}
    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []
        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])
        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }
        all_scores = []
        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])
        scores_array = np.array(all_scores)
        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)
        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })
        analysis_results[strategy_name] = analysis
    return analysis_results

def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []
    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)
    df = pd.DataFrame(df_data)
    return df, analysis_results

def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []
    print(f"Starting {num_runs} experiments...")
    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")
        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)
        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)
        print(f"Experiment {run_idx + 1} completed.")
    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df

def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "Raw Data"
    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')
    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width
    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []
    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]
        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)
    stats_df = pd.DataFrame(stats_data)
    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)
    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width
    wb.save(output_path)
    print(f"Results saved to: {output_path}")

def main():
    results_df = run_multiple_experiments(num_runs=20)
    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")
    save_results_to_excel(results_df, output_path)
    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)
    return results_df, output_path

if __name__ == "__main__":
    results, output_file = main()