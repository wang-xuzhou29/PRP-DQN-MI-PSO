import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

def execute_validation_rules(x, y, z):
    # 固定值设置: x:1--250, y:1--200, z:1--5
    triggered = set()

    # 分支1-10: 光照与温度协同控制
    if (x > 190 and x < 210 and y > 45 and y < 100) != (x > 150 and x < 210 and y > 45 and y < 100): triggered.add(1)
    if (x > 190 and x < 210 and y > 45 and y < 100) != (x > 100 and x < 210 and y > 45 and y < 100): triggered.add(2)
    if (x > 190 and x < 210 and y > 45 and y < 100) != (x > 120 and x < 210 and y > 45 and y < 100): triggered.add(3)

    if (x > 105 and x < 215 and z > 1 and z < 3) != (x > 125 and x < 215 and z > 1 and z < 3): triggered.add(4)
    if (x > 105 and x < 215 and z > 1 and z < 3) != (x > 155 and x < 215 and z > 1 and z < 3): triggered.add(5)
    if (x > 105 and x < 215 and z > 1 and z < 3) != (x > 105 and x < 245 and z > 1 and z < 3): triggered.add(6)
    if (x > 105 and x < 215 and z > 1 and z < 3) != (x > 105 and x < 215 and z > 1 and z < 4): triggered.add(7)
    if (x > 105 and x < 215 and z > 1 and z < 3) != (y > 105 and x < 215 and z > 1 and z < 3): triggered.add(8)
    if (x > 105 and x < 215 and z > 1 and z < 3) != (x > 105 or x < 215 and z > 1 and z < 3): triggered.add(9)

    if (y > 40 and y < 150 and z > 1 and z < 4) != (x > 40 and y < 150 and z > 1 and z < 4): triggered.add(10)
    if (y > 40 and y < 150 and z > 1 and z < 4) != (y > 40 and x < 150 and z > 1 and z < 4): triggered.add(11)
    if (y > 40 and y < 150 and z > 1 and z < 4) != (y > 60 and y < 150 and z > 1 and z < 4): triggered.add(12)
    if (y > 40 and y < 150 and z > 1 and z < 4) != (y > 80 and y < 150 and z > 1 and z < 4): triggered.add(13)
    if (y > 40 and y < 150 and z > 1 and z < 4) != (y > 40 and y < 170 and z > 1 and z < 4): triggered.add(14)
    if (y > 40 and y < 150 and z > 1 and z < 4) != (y > 40 and y < 150 and z > 1 or z < 4): triggered.add(15)
    if (y > 40 and y < 150 and z > 1 and z < 4) != (y > 40 and y < 150 and z > 1 and z > 4): triggered.add(16)
    if (y > 40 and y < 150 and z > 1 and z < 4) != (y > 40 and y < 150 and z > 2 and z < 4): triggered.add(17)
    if (y > 40 and y < 150 and z > 1 and z < 4) != (y > 40 and y < 150 or z > 1 and z < 4): triggered.add(18)

    if (x > 100 and x < 150) != (x > 120 and x < 150): triggered.add(19)
    if (x > 100 and x < 150) != (x > 140 and x < 150): triggered.add(20)
    if (x > 100 and x < 150) != (x > 50 and x < 150): triggered.add(21)
    if (x > 100 and x < 150) != (x > 100 and x < 200): triggered.add(22)
    if (x > 100 and x < 150) != (x > 100 and x < 210): triggered.add(23)
    if (x > 100 and x < 150) != (x < 100 and x < 150): triggered.add(24)
    if (x > 100 and x < 150) != (x > 100 and x > 150): triggered.add(25)
    if (x > 100 and x < 150) != (x > 100 and y < 150): triggered.add(26)

    if (y > 108 and y < 152) != (x > 108 and y < 152): triggered.add(27)
    if (y > 108 and y < 152) != (y > 108 and x < 152): triggered.add(28)
    if (y > 108 and y < 152) != (y > 78 and y < 152): triggered.add(29)
    if (y > 108 and y < 152) != (y > 48 and y < 152): triggered.add(30)
    if (y > 108 and y < 152) != (y > 108 and y < 142): triggered.add(31)
    if (y > 108 and y < 152) != (y > 108 and y < 132): triggered.add(32)

    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 100 > 500 and x + y + z * 100 < 700): triggered.add(33)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 100 > 600 and x + y + z * 100 < 700): triggered.add(34)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 100 > 450 and x + y + z * 100 < 700): triggered.add(35)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 100 > 400 and x + y + z * 100 < 800): triggered.add(36)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 100 > 400 and x + y + z * 100 < 750): triggered.add(37)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 100 > 400 and x + y + z * 100 < 850): triggered.add(38)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x - y + z * 100 > 400 and x + y + z * 100 < 700): triggered.add(39)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 110 > 400 and x + y + z * 100 < 700): triggered.add(40)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 120 > 400 and x + y + z * 100 < 700): triggered.add(41)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 130 > 400 and x + y + z * 100 < 700): triggered.add(42)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 100 > 400 and x + y + z * 110 < 700): triggered.add(43)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 100 > 400 and x + y + z * 120 < 700): triggered.add(44)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 100 > 400 and x + y + z * 130 < 700): triggered.add(45)
    if (x + y + z * 100 > 400 and x + y + z * 100 < 700) != (x + y + z * 100 > 400 and x - y + z * 100 < 700): triggered.add(46)

    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 50 > 85 and x * y / 100 < 115): triggered.add(47)
    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 60 > 85 and x * y / 100 < 115): triggered.add(48)
    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 80 > 85 and x * y / 100 < 115): triggered.add(49)
    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 100 > 95 and x * y / 100 < 115): triggered.add(50)
    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 100 > 105 and x * y / 100 < 115): triggered.add(51)
    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 100 > 85 or x * y / 100 < 115): triggered.add(52)
    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 100 > 85 and x * y / 100 < 125): triggered.add(53)
    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 100 > 85 and x * y / 100 < 135): triggered.add(54)
    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 100 > 85 and x * y / 100 < 155): triggered.add(55)
    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 100 > 85 and x * y / 80 < 115): triggered.add(56)
    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 100 > 85 and x * y / 70 < 115): triggered.add(57)
    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 100 > 85 and x * y / 50 < 115): triggered.add(58)
    if (x * y / 100 > 85 and x * y / 100 < 115) != (x * y / 100 > 85 and x * y / 60 < 115): triggered.add(59)

    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 180) + (z - 1) * 100 > 45 and (x - 180) + (z - 1) * 100 < 45): triggered.add(60)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 180) + (z - 1) * 100 > 35 and (x - 180) + (z - 1) * 100 < 45): triggered.add(61)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 180) + (z - 1) * 100 > 40 and (x - 180) + (z - 1) * 100 < 45): triggered.add(62)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 180) + (z - 1) * 100 > 25 or (x - 180) + (z - 1) * 100 < 45): triggered.add(63)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 80) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45): triggered.add(64)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 50) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45): triggered.add(65)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 180) + (z - 1) * 100 > 25 and (x - 80) + (z - 1) * 100 < 45): triggered.add(66)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 180) + (z - 1) * 100 > 25 and (x - 150) + (z - 1) * 100 < 45): triggered.add(67)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 180) + (z - 1) * 100 > 25 and (x - 180) - (z - 1) * 100 < 45): triggered.add(68)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 145): triggered.add(69)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 75): triggered.add(70)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 95): triggered.add(71)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 100): triggered.add(72)
    if ((x - 180) + (z - 1) * 100 > 25 and (x - 180) + (z - 1) * 100 < 45) != ((x - 180) + (z - 1) * 100 > 15 and (x - 180) + (z - 1) * 100 < 45): triggered.add(73)

    if ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145) != ((y - 50) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145): triggered.add(74)
    if ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145) != ((y + 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145): triggered.add(75)
    if ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145) != ((y - 80) + (z - 2) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145): triggered.add(76)
    if ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145) != ((y - 80) + (z - 3) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145): triggered.add(77)
    if ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145) != ((y - 80) + (z - 1) * 100 > 45 and (y - 40) + (z - 1) * 50 < 145): triggered.add(78)
    if ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145) != ((y - 80) + (z - 1) * 50 > 95 and (y - 40) + (z - 1) * 50 < 145): triggered.add(79)
    if ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145) != ((y - 80) + (z - 1) * 50 > 45 and (y - 20) + (z - 1) * 50 < 145): triggered.add(80)
    if ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145) != ((y - 80) + (z - 1) * 50 > 45 and (y + 40) + (z - 1) * 50 < 145): triggered.add(81)
    if ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145) != ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 2) * 50 < 145): triggered.add(82)
    if ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145) != ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 150 < 145): triggered.add(83)
    if ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 145) != ((y - 80) + (z - 1) * 50 > 45 and (y - 40) + (z - 1) * 50 < 200): triggered.add(84)

    if (x + y * 4 + z * 160 > 540 and x + y * 4 + z * 160 < 620) != (x + y * 5 + z * 160 > 540 and x + y * 4 + z * 160 < 620): triggered.add(85)
    if (x + y * 4 + z * 160 > 540 and x + y * 4 + z * 160 < 620) != (x + y * 6 + z * 160 > 540 and x + y * 4 + z * 160 < 620): triggered.add(86)
    if (x + y * 4 + z * 160 > 540 and x + y * 4 + z * 160 < 620) != (y + y * 4 + z * 160 > 540 and x + y * 4 + z * 160 < 620): triggered.add(87)
    if (x + y * 4 + z * 160 > 540 and x + y * 4 + z * 160 < 620) != (x + y * 4 + z * 180 > 540 and x + y * 4 + z * 160 < 620): triggered.add(88)
    if (x + y * 4 + z * 160 > 540 and x + y * 4 + z * 160 < 620) != (x + y * 4 + z * 160 > 440 and x + y * 4 + z * 160 < 620): triggered.add(89)
    if (x + y * 4 + z * 160 > 540 and x + y * 4 + z * 160 < 620) != (x + y * 4 + z * 160 > 540 and x + y * 5 + z * 160 < 620): triggered.add(90)
    if (x + y * 4 + z * 160 > 540 and x + y * 4 + z * 160 < 620) != (x + y * 4 + z * 160 > 540 and x + y * 6 + z * 160 < 620): triggered.add(91)
    if (x + y * 4 + z * 160 > 540 and x + y * 4 + z * 160 < 620) != (x + y * 4 + z * 160 > 540 and x + y * 4 + z * 160 < 720): triggered.add(92)

    if (x * y > 9500 and x * y < 15500 and z > 1 and z < 4) != (x * y > 8500 and x * y < 15500 and z > 1 and z < 4): triggered.add(93)
    if (x * y > 9500 and x * y < 15500 and z > 1 and z < 4) != (x * y > 7500 and x * y < 15500 and z > 1 and z < 4): triggered.add(94)
    if (x * y > 9500 and x * y < 15500 and z > 1 and z < 4) != (x * y > 6500 and x * y < 15500 and z > 1 and z < 4): triggered.add(95)
    if (x * y > 9500 and x * y < 15500 and z > 1 and z < 4) != (x * y > 5500 and x * y < 15500 and z > 1 and z < 4): triggered.add(96)
    if (x * y > 9500 and x * y < 15500 and z > 1 and z < 4) != (x * y > 9500 and x * y < 16500 and z > 1 and z < 4): triggered.add(97)
    if (x * y > 9500 and x * y < 15500 and z > 1 and z < 4) != (x * y > 9500 and x * y < 17500 and z > 1 and z < 4): triggered.add(98)
    if (x * y > 9500 and x * y < 15500 and z > 1 and z < 4) != (x * y > 9500 and x * y < 18500 and z > 1 and z < 4): triggered.add(99)
    if (x * y > 9500 and x * y < 15500 and z > 1 and z < 4) != (x * y > 9500 and x * y < 15500 and z > 2 and z < 4): triggered.add(100)
    if (x * y > 9500 and x * y < 15500 and z > 1 and z < 4) != (x * y > 9500 and x * y < 15500 and z > 1 and z < 5): triggered.add(101)
    if (x * y > 9500 and x * y < 15500 and z > 1 and z < 4) != (x * y > 9500 and x * y < 15500 and z > 1 or z < 4): triggered.add(102)
    if (x * y > 9500 and x * y < 15500 and z > 1 and z < 4) != (x * y > 9500 and x * y < 15500 or z > 1 and z < 4): triggered.add(103)
    if (x * y > 9500 and x * y < 15500 and z > 1 and z < 4) != (x * y > 9500 or x * y < 15500 and z > 1 and z < 4): triggered.add(104)

    if ( (x + y + z * 100) < 370 or (x + y + z * 100) > 410) != ( (x + y + z * 50) < 370 or (x + y + z * 100) > 410): triggered.add(105)
    if ( (x + y + z * 100) < 370 or (x + y + z * 100) > 410) != ( (x + y + z * 60) < 370 or (x + y + z * 100) > 410): triggered.add(106)
    if ( (x + y + z * 100) < 370 or (x + y + z * 100) > 410) != ( (x + y + z * 70) < 370 or (x + y + z * 100) > 410): triggered.add(107)
    if ( (x + y + z * 100) < 370 or (x + y + z * 100) > 410) != ( (x + y + z * 80) < 370 or (x + y + z * 100) > 410): triggered.add(108)
    if ( (x + y + z * 100) < 370 or (x + y + z * 100) > 410) != ( (x + y + z * 100) < 270 or (x + y + z * 100) > 410): triggered.add(109)
    if ( (x + y + z * 100) < 370 or (x + y + z * 100) > 410) != ( (x + y + z * 100) < 370 or (x + y + z * 100) > 510): triggered.add(110)
    if ( (x + y + z * 100) < 370 or (x + y + z * 100) > 410) != ( (x + y + z * 100) < 370 or (x + y + z * 100) > 610): triggered.add(111)
    if ( (x + y + z * 100) < 370 or (x + y + z * 100) > 410) != ( (x + y + z * 100) < 370 or (x + y + z * 100) > 710): triggered.add(112)
    if ( (x + y + z * 100) < 370 or (x + y + z * 100) > 410) != ( (x + y + z * 100) < 370 or (x + y + z * 50) > 410): triggered.add(113)
    if ( (x + y + z * 100) < 370 or (x + y + z * 100) > 410) != ( (x + y + z * 100) < 370 or (x + y + z * 70) > 410): triggered.add(114)
    if ( (x + y + z * 100) < 370 or (x + y + z * 100) > 410) != ( (x + x + z * 100) < 370 or (x + y + z * 100) > 410): triggered.add(115)
    if ( (x + y + z * 100) < 370 or (x + y + z * 100) > 410) != ( (x + y + z * 100) < 370 or (x + x + z * 100) > 410): triggered.add(116)

    return triggered

def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """"""
    generated_path = execute_validation_rules(particle[0], particle[1], particle[2])

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # : x:1-50, y:1-50, z:1-50
        self.bounds = bounds if bounds else [(1, 250), (1, 200), (1, 5)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """target pathsPSO"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(particles[i][0], particles[i][1], particles[i][2]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(gbest_particle[0], gbest_particle[1], gbest_particle[2]),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """Path PSO"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - Path ")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations")
    print(f"Path : {len(target_paths)}")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"Path {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "" if result['success'] else f"({result['best_fitness']:.3f})"
        print(f"{status} | {result['time']:.2f}s | iterations{result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f": {success_count}/{len(target_paths)} ({success_rate:.1f}%) | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs}")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations, {len(target_paths)}Path ")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"---  {run_idx}/{num_runs} ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f": {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs} runcompleted | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 1: 
    ws1 = wb.active
    ws1.title = ""
    ws1.sheet_view.showGridLines = False

    headers = ["Run", "", "", "", "Average Iterations", "(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f" {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # 2: Path 
    ws2 = wb.create_sheet(title="Path ")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Path ID", "", "", "", "Average Iterations", "Minimum Iterations", "Maximum Iterations"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"Path  {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # 3: 
    ws3 = wb.create_sheet(title="")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Path ", "", "(x,y,z)", "", "Iterations", "Path "]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"{run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # 4: target paths
    ws4 = wb.create_sheet(title="target paths")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Path ID", "target paths", ""]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"Path  {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f" : {filename}")
    print(f"{'=' * 70}")
    print(f":")
    print(f"  1.        - {len(all_results)} run")
    print(f"  2. Path        - Path ")
    print(f"  3.    -  runPath ")
    print(f"  4. target paths       - target paths")
    print(f"{'=' * 70}\n")

    return filename


def main():


    target_paths = [
        {1, 2, 3, 5, 8, 11, 12, 13, 16, 17, 22, 23, 25, 26, 27, 30, 33, 34, 35, 39, 50, 51, 57, 58, 59, 63, 68, 69, 70,
         71,
         72, 74, 75, 78, 92, 93, 94, 95, 96, 102, 103, 104, 110, 111, 112, 113, 114},

        {1, 2, 3, 8, 11, 12, 13, 16, 17, 22, 23, 25, 26, 27, 30, 33, 34, 35, 39, 50, 51, 57, 58, 59, 63, 68, 69, 71, 72,
         74,
         75, 78, 92, 93, 94, 95, 96, 102, 103, 104, 110, 111, 112, 113, 114},

        {2, 4, 5, 8, 13, 16, 17, 19, 20, 24, 25, 27, 29, 30, 40, 41, 42, 50, 51, 56, 57, 58, 59, 60, 62, 66, 67, 76, 77,
         79,
         81, 83, 93, 94, 95, 96, 102, 103, 104, 105, 106, 107, 108, 116},

        {2, 4, 5, 8, 16, 17, 19, 20, 24, 25, 27, 29, 30, 40, 41, 42, 50, 51, 56, 57, 58, 59, 60, 61, 62, 66, 67, 76, 77,
         79,
         81, 83, 93, 94, 95, 96, 102, 103, 104, 105, 106, 107, 108, 116},

        {2, 3, 5, 8, 11, 12, 13, 16, 17, 22, 23, 26, 27, 30, 33, 34, 35, 39, 47, 48, 49, 52, 63, 68, 69, 70, 71, 72, 74,
         75,
         78, 92, 94, 95, 96, 102, 103, 104, 105, 106, 107, 108, 116},

        {2, 7, 9, 13, 16, 19, 20, 24, 25, 27, 29, 30, 33, 34, 39, 50, 51, 56, 57, 58, 59, 63, 68, 69, 77, 80, 81, 83,
         93,
         94, 95, 96, 102, 103, 104, 110, 111, 112, 113, 114},

        {2, 3, 4, 5, 8, 13, 16, 17, 20, 24, 25, 27, 29, 30, 33, 34, 35, 39, 51, 56, 57, 58, 59, 63, 68, 69, 70, 71, 72,
         76,
         77, 79, 81, 83, 100, 105, 106, 107, 108, 116},

        {1, 2, 3, 5, 8, 11, 13, 16, 17, 22, 23, 25, 26, 27, 30, 33, 34, 35, 39, 52, 53, 54, 55, 63, 68, 69, 70, 71, 72,
         76,
         77, 79, 81, 83, 100, 110, 111, 112, 113, 114},

        {1, 2, 3, 8, 11, 16, 17, 22, 23, 25, 26, 27, 29, 30, 33, 34, 39, 52, 63, 68, 69, 72, 76, 77, 79, 81, 83, 97, 98,
         99,
         102, 103, 104, 110, 111, 112, 113, 114},

        {8, 9, 11, 15, 18, 21, 24, 28, 33, 34, 35, 39, 50, 51, 56, 57, 58, 59, 63, 64, 65, 82, 84, 93, 94, 95, 96, 102,
         103,
         104, 110, 111, 112, 113, 114, 116},
        {9, 16, 17, 21, 24, 29, 30, 40, 41, 42, 50, 51, 56, 57, 58, 59, 63, 64, 65, 73, 76, 77, 79, 81, 83, 93, 94, 95,
         96,
         102, 103, 104, 105, 106, 107, 108},
        {6, 9, 11, 12, 13, 16, 17, 25, 26, 27, 33, 34, 39, 50, 51, 57, 58, 59, 63, 68, 69, 75, 78, 92, 93, 94, 95, 96,
         102,
         103, 104, 110, 111, 112, 113, 114},
        {8, 9, 16, 17, 21, 24, 27, 40, 41, 42, 50, 51, 57, 58, 59, 63, 64, 65, 76, 77, 79, 81, 83, 93, 94, 95, 96, 102,
         103,
         104, 105, 106, 107, 108, 115},
        {8, 9, 16, 17, 21, 24, 27, 32, 33, 34, 35, 39, 50, 51, 56, 57, 58, 59, 63, 64, 65, 82, 84, 93, 94, 95, 96, 102,
         103,
         104, 105, 106, 107, 108, 115},
        {5, 8, 10, 15, 18, 22, 23, 25, 26, 27, 40, 41, 42, 47, 48, 52, 63, 68, 69, 70, 71, 72, 75, 78, 87, 90, 91, 102,
         103,
         104, 105, 106, 107, 108, 116},
        {8, 9, 11, 14, 15, 18, 21, 24, 28, 33, 34, 35, 39, 47, 48, 49, 52, 63, 64, 65, 82, 84, 94, 95, 96, 102, 103,
         104,
         110, 111, 112, 113, 114, 116},
        {2, 4, 5, 8, 12, 13, 16, 17, 19, 20, 24, 25, 27, 30, 41, 42, 47, 48, 52, 60, 61, 62, 66, 67, 74, 75, 78, 92, 96,
         102, 103, 104, 109, 115},
        {4, 5, 16, 17, 19, 20, 24, 25, 31, 32, 33, 34, 39, 52, 60, 61, 62, 66, 67, 82, 84, 97, 98, 99, 102, 103, 104,
         110,
         111, 112, 113, 114},
        {5, 8, 10, 15, 18, 22, 23, 25, 26, 27, 41, 42, 52, 63, 68, 69, 71, 72, 75, 85, 86, 88, 89, 102, 103, 104, 105,
         106,
         107, 108, 116},
        {1, 2, 3, 9, 15, 18, 22, 23, 25, 26, 27, 29, 30, 44, 45, 52, 53, 54, 55, 63, 68, 82, 84, 101, 102, 103, 104,
         112},
        {1, 2, 3, 9, 15, 16, 18, 22, 23, 25, 26, 27, 29, 30, 36, 37, 38, 46, 52, 53, 54, 55, 63, 68, 102, 103, 104},
        {2, 9, 15, 16, 18, 19, 20, 24, 25, 27, 29, 30, 43, 44, 45, 51, 56, 57, 58, 59, 63, 68, 102, 103, 104, 112},
    ]

    print("=" * 70)
    print("baseline PSO")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("Program completed")


if __name__ == "__main__":
    main()