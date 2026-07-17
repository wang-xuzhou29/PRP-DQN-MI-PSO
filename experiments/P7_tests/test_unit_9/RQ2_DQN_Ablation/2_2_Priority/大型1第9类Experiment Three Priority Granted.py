import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import math

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# ========== 新的取值范围（浮点数，对应 section9） ==========
X_MIN, X_MAX = 140.0, 240.0   # 速度范围
Y_MIN, Y_MAX = 200.0, 250.0   # 扭矩范围
Z_MIN, Z_MAX = 1.0, 90.0      # 电流范围
# ==========================================================

# === 归一化、反归一化、有效性检查、裁剪（支持浮点数） ===
def normalize_state(state):
    """
    将 (x,y,z) 归一化到 [0,1] 范围
    """
    x, y, z = state
    norm_x = (x - X_MIN) / (X_MAX - X_MIN)
    norm_y = (y - Y_MIN) / (Y_MAX - Y_MIN)
    norm_z = (z - Z_MIN) / (Z_MAX - Z_MIN)
    return (norm_x, norm_y, norm_z)


def denormalize_state(norm_state):
    """
    将 [0,1] 归一化值还原为原始浮点数坐标（保留浮点数）
    """
    norm_x, norm_y, norm_z = norm_state
    x = norm_x * (X_MAX - X_MIN) + X_MIN
    y = norm_y * (Y_MAX - Y_MIN) + Y_MIN
    z = norm_z * (Z_MAX - Z_MIN) + Z_MIN
    # 裁剪确保在范围内（浮点数）
    x = max(X_MIN, min(X_MAX, x))
    y = max(Y_MIN, min(Y_MAX, y))
    z = max(Z_MIN, min(Z_MAX, z))
    return (x, y, z)


def is_valid_state(state):
    """检查状态是否在有效范围内（浮点数）"""
    x, y, z = state
    return (X_MIN <= x <= X_MAX and
            Y_MIN <= y <= Y_MAX and
            Z_MIN <= z <= Z_MAX)


def clip_state(state):
    """将状态裁剪到有效范围（浮点数）"""
    x, y, z = state
    return (
        max(X_MIN, min(X_MAX, x)),
        max(Y_MIN, min(Y_MAX, y)),
        max(Z_MIN, min(Z_MAX, z))
    )


# === 指标收集器（优先级经验回放专用） ===
class PrioritizedMetricsCollector:
    def __init__(self, experiment_name="Prioritized_DQN"):
        self.experiment_name = experiment_name
        self.start_time = None
        self.end_time = None

        # 基础指标
        self.total_reward = 0
        self.td_errors = []
        self.final_output_similarities = []
        self.action_improvements = []
        self.total_memory_usage = 0
        self.memory_check_count = 0
        self.step_count = 0

        # 每回合指标
        self.episode_rewards = []
        self.episode_similarities = []
        self.episode_td_errors = []
        self.episode_epsilon_values = []
        self.episode_memory_usage = []

        # 优先级相关指标
        self.priority_statistics = []
        self.importance_weights = []
        self.high_priority_samples_ratio = []
        self.priority_distribution_stats = []

        # 分组路径性能
        self.similar_paths_performance = []
        self.isolated_paths_performance = []

        # 里程碑数据
        self.milestone_data = {}

        # 收敛检测
        self.convergence_window = 20
        self.convergence_threshold = 0.02
        self.convergence_detected_episode = None

        # 样本效率数据
        self.sample_efficiency_data = []
        self.performance_milestones = [0.6, 0.7, 0.75, 0.8]

        # 学习曲线特性
        self.learning_curve_characteristics = {}
        self.early_vs_late_performance = {}

    def reset(self):
        """重置所有指标，用于新一次运行"""
        self.start_time = None
        self.end_time = None
        self.total_reward = 0
        self.td_errors = []
        self.final_output_similarities = []
        self.action_improvements = []
        self.total_memory_usage = 0
        self.memory_check_count = 0
        self.step_count = 0
        self.episode_rewards = []
        self.episode_similarities = []
        self.episode_td_errors = []
        self.episode_epsilon_values = []
        self.episode_memory_usage = []
        self.priority_statistics = []
        self.importance_weights = []
        self.high_priority_samples_ratio = []
        self.priority_distribution_stats = []
        self.similar_paths_performance = []
        self.isolated_paths_performance = []
        self.milestone_data = {}
        self.convergence_detected_episode = None
        self.sample_efficiency_data = []
        self.learning_curve_characteristics = {}
        self.early_vs_late_performance = {}

    def start_training(self):
        self.start_time = time.time()

    def end_training(self):
        self.end_time = time.time()

    def record_step_metrics(self, reward, td_error, triggered, target_path, priority=None, is_weight=None):
        """记录每一步的指标"""
        self.step_count += 1
        self.total_reward += reward
        self.td_errors.append(td_error)

        if priority is not None:
            self.priority_statistics.append(priority)
        if is_weight is not None:
            self.importance_weights.append(is_weight)

        process = psutil.Process(os.getpid())
        current_memory = process.memory_info().rss / 1024 / 1024
        self.total_memory_usage += current_memory
        self.memory_check_count += 1

    def record_episode_metrics(self, episode, episode_reward, avg_similarity, avg_td_error, epsilon,
                               path_group="similar", priority_stats=None):
        """记录每个回合的指标"""
        self.episode_rewards.append(episode_reward)
        self.episode_similarities.append(avg_similarity)
        self.episode_td_errors.append(avg_td_error)
        self.episode_epsilon_values.append(epsilon)

        process = psutil.Process(os.getpid())
        current_memory = process.memory_info().rss / 1024 / 1024
        self.episode_memory_usage.append(current_memory)

        if priority_stats:
            self.priority_distribution_stats.append({
                'episode': episode,
                'mean_priority': priority_stats.get('mean_priority', 0),
                'max_priority': priority_stats.get('max_priority', 0),
                'min_priority': priority_stats.get('min_priority', 0),
                'high_priority_ratio': priority_stats.get('high_priority_ratio', 0)
            })

        if path_group == "similar":
            self.similar_paths_performance.append({
                'episode': episode,
                'reward': episode_reward,
                'similarity': avg_similarity,
                'td_error': avg_td_error
            })
        else:
            self.isolated_paths_performance.append({
                'episode': episode,
                'reward': episode_reward,
                'similarity': avg_similarity,
                'td_error': avg_td_error
            })

        if episode in [50, 100, 150, 200, 250, 300, 400, 450, 500]:
            self.milestone_data[episode] = {
                'avg_reward': np.mean(self.episode_rewards[-10:]) if len(self.episode_rewards) >= 10 else episode_reward,
                'avg_similarity': avg_similarity,
                'avg_td_error': avg_td_error,
                'epsilon': epsilon,
                'memory_usage': current_memory,
                'total_steps': self.step_count,
                'priority_stats': priority_stats
            }

        self._check_convergence(episode)
        self._check_performance_milestones(episode, avg_similarity)

    def _check_convergence(self, episode):
        """检测是否收敛"""
        if len(self.episode_similarities) >= self.convergence_window and self.convergence_detected_episode is None:
            recent_similarities = self.episode_similarities[-self.convergence_window:]
            if np.std(recent_similarities) < self.convergence_threshold:
                self.convergence_detected_episode = episode

    def _check_performance_milestones(self, episode, similarity):
        """检测是否达到性能里程碑"""
        for milestone in self.performance_milestones:
            if similarity >= milestone and not any(data[1] == milestone for data in self.sample_efficiency_data):
                self.sample_efficiency_data.append((episode, milestone, self.step_count))

    def record_final_output_sample(self, triggered, target_path):
        """记录最终输出样本的相似度"""
        if len(triggered | target_path) > 0:
            similarity = len(triggered & target_path) / len(triggered | target_path)
        else:
            similarity = 0.0
        self.final_output_similarities.append(similarity)

    def record_action_improvement(self, current_reward, prev_reward):
        """记录动作改进情况"""
        if prev_reward is not None:
            improvement = current_reward - prev_reward
            self.action_improvements.append(1 if improvement > 0 else 0)


# 全局指标收集器实例
prioritized_metrics = PrioritizedMetricsCollector("Prioritized_DQN_Enhanced")


# === 奖励函数 ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward


# ========================== section9 函数（速度扭矩电流混合，新范围） ==========================
def section9_hybrid_speed_torque_current(x, y, z):
    """第9类: 速度扭矩电流混合控制 (速度x, 扭矩y, 电流z) - 160个有效变异分支"""
    triggered = set()
    # 标准范围与状态范围一致
    speed_std = (140, 240)
    torque_std = (200, 250)
    current_std = (1, 90)

    if [(160 < x < 240)] != [(160 < x * 8 < 240)]:
        triggered.add(1)
    if [(160 < x < 240)] != [(160 < 100 < 240)]:
        triggered.add(2)

    if [(210 < y < 250)] != [(210 < y * 9 < 250)]:
        triggered.add(3)
    if [(210 < y < 250)] != [(210 < y * 12 < 250)]:
        triggered.add(4)

    if [(40 < z < 90)] != [(40 < z * 8 < 90)]:
        triggered.add(5)
    if [(40 < z < 90)] != [(40 < z * 9 < 90)]:
        triggered.add(6)

    if [(x > 190 and x < 210 and y > 225 and y < 235)] != [(x > 190 and x < 210 and y > 225 and 200 < 235)]:
        triggered.add(7)
    if [(x > 190 and x < 210 and y > 225 and y < 235)] != [(x > 190 and x < 210 and y > 225 and 220 < 235)]:
        triggered.add(8)

    if [(x > 185 and x < 215 and z > 60 and z < 70)] != [(x > 185 and x < 215 and z > 60 and 60 < 70)]:
        triggered.add(9)
    if [(x > 185 and x < 215 and z > 60 and z < 70)] != [(x > 185 and x < 215 and z > 60 and 55 < 70)]:
        triggered.add(10)

    if [(y > 223 and y < 237 and z > 60 and z < 70)] != [(y > 223 and y < 237 and z > 60 and 60 < 70)]:
        triggered.add(11)
    if [(y > 223 and y < 237 and z > 60 and z < 70)] != [(y > 223 and y < 237 and z > 60 and 67 < 70)]:
        triggered.add(12)

    if [(x > 198 and x < 202)] != [(x > 198 and x * 8 < 202)]:
        triggered.add(13)
    if [(x > 198 and x < 202)] != [(x > 198 and 200 < 202)]:
        triggered.add(14)

    if [(y > 228 and y < 232)] != [(y > 228 and y * 8 < 232)]:
        triggered.add(15)
    if [(y > 228 and y < 232)] != [(y > 228 and 200 < 232)]:
        triggered.add(16)

    if [(z > 63 and z < 67)] != [(z > 63 and 44 < 67)]:
        triggered.add(17)
    if [(z > 63 and z < 67)] != [(z > 63 and 56 < 67)]:
        triggered.add(18)

    if [(abs(x - 200) < 5)] != [(abs(x - 200) < 51)]:
        triggered.add(19)
    if [(abs(x - 200) < 5)] != [(abs(x - 200) < 15)]:
        triggered.add(20)

    if [(abs(y - 230) < 3)] != [(abs(y - 230) < 13)]:
        triggered.add(21)
    if [(abs(y - 230) < 3)] != [(abs(y - 230) < 31)]:
        triggered.add(22)

    if [(abs(z - 65) < 2)] != [(abs(z - 65) < 21)]:
        triggered.add(23)
    if [(abs(z - 65) < 2)] != [(abs(z - 65) < 12)]:
        triggered.add(24)

    if [(x > 185 and y > 223 and z > 55)] != [(x > 185 and y > 223 and z > 155)]:
        triggered.add(25)
    if [(x > 185 and y > 223 and z > 55)] != [(x > 185 and y > 223 and z > 515)]:
        triggered.add(26)

    if [(x < 215 and y < 237 and z < 75)] != [(x < 215 and y < 237 and z < 715)]:
        triggered.add(27)
    if [(x < 215 and y < 237 and z < 75)] != [(x < 215 and y < 237 and z < 225)]:
        triggered.add(28)

    if [((x + y / 10 + z) > 288 and (x + y / 10 + z) < 308)] != [((x + y / 10 + z) > 288 and (x + y / 10 + z) < 508)]:
        triggered.add(29)
    if [((x + y / 10 + z) > 288 and (x + y / 10 + z) < 308)] != [((x + y / 10 + z) > 288 and (x + y / 10 + z) < 708)]:
        triggered.add(30)

    if [(x * y / 100 > 440 and x * y / 100 < 500)] != [(x * y / 100 > 440 and x * y / 100 < 900)]:
        triggered.add(31)
    if [(x * y / 100 > 440 and x * y / 100 < 500)] != [(x * y / 100 > 440 and x * y / 100 < 440)]:
        triggered.add(32)

    if [(x * z > 12500 and x * z < 14500)] != [(x * z > 12500 and x * z < 17600)]:
        triggered.add(33)
    if [(x * z > 12500 and x * z < 14500)] != [(x * z > 12500 and x * z < 54500)]:
        triggered.add(34)

    if [(y / 10 * z > 1400 and y / 10 * z < 1600)] != [(y / 10 * z > 1400 and y / 10 * z < 7800)]:
        triggered.add(35)
    if [(y / 10 * z > 1400 and y / 10 * z < 1600)] != [(y / 10 * z > 1400 and y / 10 * z < 5600)]:
        triggered.add(36)

    if [((x + y / 10 + z) / 3 > 96 and (x + y / 10 + z) / 3 < 102)] != [((x + y / 10 + z) / 3 > 96 and (x + y / 10 + z) / 3 < 1052)]:
        triggered.add(37)
    if [((x + y / 10 + z) / 3 > 96 and (x + y / 10 + z) / 3 < 102)] != [((x + y / 10 + z) / 3 > 96 and (x + y / 10 + z) / 3 < 1502)]:
        triggered.add(38)

    if [(math.sqrt(x ** 2 + (y / 10) ** 2 + z ** 2) > 220)] != [(math.sqrt(x ** 2 + (y / 10) ** 2 + z ** 2) > 2520)]:
        triggered.add(39)
    if [(math.sqrt(x ** 2 + (y / 10) ** 2 + z ** 2) > 220)] != [(math.sqrt(x ** 2 + (y / 10) ** 2 + z ** 2) > 280)]:
        triggered.add(40)

    if [(x / (y / 10) > 8.2 and x / (y / 10) < 9.2)] != [(x / (y / 10) > 8.2 and x / (y / 10) < 92)]:
        triggered.add(41)
    if [(x / (y / 10) > 8.2 and x / (y / 10) < 9.2)] != [(x / (y / 10) > 8.2 and x / (y / 10) < 19.2)]:
        triggered.add(42)

    if [(x / z > 2.8 and x / z < 3.4)] != [(x / z > 2.8 and x / z < 13.4)]:
        triggered.add(43)
    if [(x / z > 2.8 and x / z < 3.4)] != [(x / z > 2.8 and x / z < 31.4)]:
        triggered.add(44)

    if [((y / 10) / z > 0.32 and (y / 10) / z < 0.38)] != [((y / 10) / z > 0.32 and (y / 10) / z < 38)]:
        triggered.add(45)
    if [((y / 10) / z > 0.32 and (y / 10) / z < 0.38)] != [((y / 10) / z > 0.32 and (y / 10) / z < 3.8)]:
        triggered.add(46)

    if [((x - 180) + (y - 220) / 10 > 22 and (x - 180) + (y - 220) / 10 < 32)] != [((x - 180) + (y - 220) / 10 > 22 and (x - 180) + (y - 220) / 10 < 312)]:
        triggered.add(47)
    if [((x - 180) + (y - 220) / 10 > 22 and (x - 180) + (y - 220) / 10 < 32)] != [((x - 180) + (y - 220) / 10 > 22 and (x - 180) + (y - 220) / 10 < 132)]:
        triggered.add(48)

    if [((x - 180) + (z - 50) > 35 and (x - 180) + (z - 50) < 45)] != [((x - 180) + (z - 50) > 35 and (x - 180) + (z - 50) < 415)]:
        triggered.add(49)
    if [((x - 180) + (z - 50) > 35 and (x - 180) + (z - 50) < 45)] != [((x - 180) + (z - 50) > 35 and (x - 180) + (z - 50) < 145)]:
        triggered.add(50)

    if [((y - 220) / 10 + (z - 50) > 25 and (y - 220) / 10 + (z - 50) < 35)] != [((y - 220) / 10 + (z - 50) > 25 and (y - 220) / 10 + (z - 50) < 315)]:
        triggered.add(51)
    if [((y - 220) / 10 + (z - 50) > 25 and (y - 220) / 10 + (z - 50) < 35)] != [((y - 220) / 10 + (z - 50) > 25 and (y - 220) / 10 + (z - 50) < 135)]:
        triggered.add(52)

    if [(abs((x - 200) - (y - 230) / 2) < 8)] != [(abs((x - 200) - (y - 230) / 2) < 18)]:
        triggered.add(53)
    if [(abs((x - 200) - (y - 230) / 2) < 8)] != [(abs((x - 200) - (y - 230) / 2) < 81)]:
        triggered.add(54)

    if [(abs((x - 200) - (z - 65) * 3) < 12)] != [(abs((x - 200) - (z - 65) * 3) < 92)]:
        triggered.add(55)
    if [(abs((x - 200) - (z - 65) * 3) < 12)] != [(abs((x - 200) - (z - 65) * 3) < 122)]:
        triggered.add(56)

    if [(abs((y - 230) / 10 - (z - 65) / 5) < 3)] != [(abs((y - 230) / 10 - (z - 65) / 5) < 13)]:
        triggered.add(57)
    if [(abs((y - 230) / 10 - (z - 65) / 5) < 3)] != [(abs((y - 230) / 10 - (z - 65) / 5) < 31)]:
        triggered.add(58)

    if [(x / (y / 10 + 5) > 6.5 and x / (y / 10 + 5) < 8.5)] != [(x / (y / 10 + 5) > 6.5 and x / (y / 10 + 5) < 85)]:
        triggered.add(59)
    if [(x / (y / 10 + 5) > 6.5 and x / (y / 10 + 5) < 8.5)] != [(x / (y / 10 + 5) > 6.5 and x / (y / 10 + 5) < 81.5)]:
        triggered.add(60)

    if [((y / 10) / (z + 15) > 0.26 and (y / 10) / (z + 15) < 0.34)] != [((y / 10) / (z + 15) > 0.26 and (y / 10) / (z + 15) < 10.34)]:
        triggered.add(61)
    if [((y / 10) / (z + 15) > 0.26 and (y / 10) / (z + 15) < 0.34)] != [((y / 10) / (z + 15) > 0.26 and (y / 10) / (z + 15) < 40.34)]:
        triggered.add(62)

    if [(z / (x / 5) > 1.4 and z / (x / 5) < 1.8)] != [(z / (x / 5) > 1.4 and z / (x / 5) < 11.8)]:
        triggered.add(63)
    if [(z / (x / 5) > 1.4 and z / (x / 5) < 1.8)] != [(z / (x / 5) > 1.4 and z / (x / 5) < 41.8)]:
        triggered.add(64)

    if [((x * y * z / 10000) > 28 and (x * y * z / 10000) < 36)] != [((x * y * z / 10000) > 28 and (x * y * z / 10000) < 332)]:
        triggered.add(65)
    if [((x * y * z / 10000) > 28 and (x * y * z / 10000) < 36)] != [((x * y * z / 10000) > 28 and (x * y * z / 10000) < 236)]:
        triggered.add(66)

    if [(x * 0.5 + y / 10 * 0.3 + z * 0.2 > 125 and x * 0.5 + y / 10 * 0.3 + z * 0.2 < 135)] != [(x * 0.5 + y / 10 * 0.3 + z * 0.2 > 125 and x * 0.5 + y / 10 * 0.3 + z * 0.2 < 935)]:
        triggered.add(67)
    if [(x * 0.5 + y / 10 * 0.3 + z * 0.2 > 125 and x * 0.5 + y / 10 * 0.3 + z * 0.2 < 135)] != [(x * 0.5 + y / 10 * 0.3 + z * 0.2 > 125 and x * 0.5 + y / 10 * 0.3 + z * 0.2 < 635)]:
        triggered.add(68)

    if [((x / 200) ** 0.4 * (y / 230) ** 0.3 * (z / 65) ** 0.3 > 0.92)] != [((x / 200) ** 0.4 * (y / 230) ** 0.3 * (z / 65) ** 0.3 > 9.2)]:
        triggered.add(69)
    if [((x / 200) ** 0.4 * (y / 230) ** 0.3 * (z / 65) ** 0.3 > 0.92)] != [((x / 200) ** 0.4 * (y / 230) ** 0.3 * (z / 65) ** 0.3 > 92)]:
        triggered.add(70)

    if [((x - 200) * (y - 230) / 10 > -80 and (x - 200) * (y - 230) / 10 < 80)] != [((x - 200) * (y - 230) / 10 > -80 and (x - 200) * (y - 230) / 10 < 180)]:
        triggered.add(71)
    if [((x - 200) * (y - 230) / 10 > -80 and (x - 200) * (y - 230) / 10 < 80)] != [((x - 200) * (y - 230) / 10 > -80 and (x - 200) * (y - 230) / 10 < 810)]:
        triggered.add(72)

    if [((x - 200) * (z - 65) > -120 and (x - 200) * (z - 65) < 120)] != [((x - 200) * (z - 65) > -120 and (x - 200) * (z - 65) < 820)]:
        triggered.add(73)
    if [((x - 200) * (z - 65) > -120 and (x - 200) * (z - 65) < 120)] != [((x - 200) * (z - 65) > -120 and (x - 200) * (z - 65) < 520)]:
        triggered.add(74)

    if [((y - 230) / 10 * (z - 65) > -30 and (y - 230) / 10 * (z - 65) < 30)] != [((y - 230) / 10 * (z - 65) > -30 and (y - 230) / 10 * (z - 65) < 310)]:
        triggered.add(75)
    if [((y - 230) / 10 * (z - 65) > -30 and (y - 230) / 10 * (z - 65) < 30)] != [((y - 230) / 10 * (z - 65) > -30 and (y - 230) / 10 * (z - 65) < 130)]:
        triggered.add(76)

    if [(abs(x / 200 + y / 230 + z / 65 - 3) < 0.18)] != [(abs(x / 200 + y / 230 + z / 65 - 3) < 18)]:
        triggered.add(77)
    if [(abs(x / 200 + y / 230 + z / 65 - 3) < 0.18)] != [(abs(x / 200 + y / 230 + z / 65 - 3) < 10.18)]:
        triggered.add(78)

    if [(abs((x / 200) * (y / 230) * (z / 65) - 1) < 0.12)] != [(abs((x / 200) * (y / 230) * (z / 65) - 1) < 10.12)]:
        triggered.add(79)
    if [(abs((x / 200) * (y / 230) * (z / 65) - 1) < 0.12)] != [(abs((x / 200) * (y / 230) * (z / 65) - 1) < 12.12)]:
        triggered.add(80)

    if [(x > 197 and x < 203 and y > 228 and y < 232 and z > 63.5 and z < 66.5)] != [(x > 197 and x < 203 and y > 228 and y < 232 and z > 63.5 and z < 616.5)]:
        triggered.add(81)
    if [(x > 197 and x < 203 and y > 228 and y < 232 and z > 63.5 and z < 66.5)] != [(x > 197 and x < 203 and y > 228 and y < 232 and z > 63.5 and z < 166.5)]:
        triggered.add(82)

    if [(abs(x / (y / 10) - 8.7) < 0.3 and abs(z - 65) < 1.5)] != [(abs(x / (y / 10) - 8.7) < 0.3 and abs(z - 65) < 11.5)]:
        triggered.add(83)
    if [(abs(x / (y / 10) - 8.7) < 0.3 and abs(z - 65) < 1.5)] != [(abs(x / (y / 10) - 8.7) < 0.3 and abs(z - 65) < 15)]:
        triggered.add(84)

    if [(x * y / 100 > 455 and x * y / 100 < 465 and z > 64 and z < 66)] != [(x * y / 100 > 455 and x * y / 100 < 465 and z > 64 and z < 166)]:
        triggered.add(85)
    if [(x * y / 100 > 455 and x * y / 100 < 465 and z > 64 and z < 66)] != [(x * y / 100 > 455 and x * y / 100 < 465 and z > 64 and z < 616)]:
        triggered.add(86)

    if [((x + y / 10) / 2 > 112 and (x + y / 10) / 2 < 116 and z > 63.5 and z < 66.5)] != [((x + y / 10) / 2 > 112 and (x + y / 10) / 2 < 116 and z > 63.5 and z < 166.5)]:
        triggered.add(87)
    if [((x + y / 10) / 2 > 112 and (x + y / 10) / 2 < 116 and z > 63.5 and z < 66.5)] != [((x + y / 10) / 2 > 112 and (x + y / 10) / 2 < 116 and z > 63.5 and z < 661.5)]:
        triggered.add(88)

    if [(abs(x - y / 10) < 175 and z > 63.5 and z < 66.5)] != [(abs(x - y / 10) < 175 and z > 63.5 and z < 616.5)]:
        triggered.add(89)
    if [(abs(x - y / 10) < 175 and z > 63.5 and z < 66.5)] != [(abs(x - y / 10) < 175 and z > 63.5 and z < 6615)]:
        triggered.add(90)

    if [(math.sqrt((x - 200) ** 2 + (y / 10 - 23) ** 2) < 3 and abs(z - 65) < 1.5)] != [(math.sqrt((x - 200) ** 2 + (y / 10 - 23) ** 2) < 3 and abs(z - 65) < 15)]:
        triggered.add(91)
    if [(math.sqrt((x - 200) ** 2 + (y / 10 - 23) ** 2) < 3 and abs(z - 65) < 1.5)] != [(math.sqrt((x - 200) ** 2 + (y / 10 - 23) ** 2) < 3 and abs(z - 65) < 11.5)]:
        triggered.add(92)

    if [(x / 200 > 0.985 and x / 200 < 1.015 and y / 230 > 0.987 and y / 230 < 1.013)] != [(x / 200 > 0.985 and x / 200 < 1.015 and y / 230 > 0.987 and y / 230 < 10.13)]:
        triggered.add(93)
    if [(x / 200 > 0.985 and x / 200 < 1.015 and y / 230 > 0.987 and y / 230 < 1.013)] != [(x / 200 > 0.985 and x / 200 < 1.015 and y / 230 > 0.987 and y / 230 < 1013)]:
        triggered.add(94)

    if [(z / 65 > 0.985 and z / 65 < 1.015)] != [(z / 65 > 0.985 and z / 65 < 121.015)]:
        triggered.add(95)
    if [(z / 65 > 0.985 and z / 65 < 1.015)] != [(z / 65 > 0.985 and z / 65 < 111.015)]:
        triggered.add(96)

    if [((x / 200 + y / 230 + z / 65) / 3 > 0.985 and (x / 200 + y / 230 + z / 65) / 3 < 1.015)] != [((x / 200 + y / 230 + z / 65) / 3 > 0.985 and (x / 200 + y / 230 + z / 65) / 3 < 2.115)]:
        triggered.add(97)
    if [((x / 200 + y / 230 + z / 65) / 3 > 0.985 and (x / 200 + y / 230 + z / 65) / 3 < 1.015)] != [((x / 200 + y / 230 + z / 65) / 3 > 0.985 and (x / 200 + y / 230 + z / 65) / 3 < 15)]:
        triggered.add(98)

    if [(max(abs(x / 200 - 1), abs(y / 230 - 1), abs(z / 65 - 1)) < 0.02)] != [(max(abs(x / 200 - 1), abs(y / 230 - 1), abs(z / 65 - 1)) < 22)]:
        triggered.add(99)
    if [(max(abs(x / 200 - 1), abs(y / 230 - 1), abs(z / 65 - 1)) < 0.02)] != [(max(abs(x / 200 - 1), abs(y / 230 - 1), abs(z / 65 - 1)) < 10.02)]:
        triggered.add(100)

    if [(min(x / 200, y / 230, z / 65) > 0.985)] != [(min(x / 200, y / 230, z / 65) > 9.85)]:
        triggered.add(101)
    if [(min(x / 200, y / 230, z / 65) > 0.985)] != [(min(x / 200, y / 230, z / 65) > 10.985)]:
        triggered.add(102)

    if [(max(x / 200, y / 230, z / 65) < 1.015)] != [(max(x / 200, y / 230, z / 65) < 10.15)]:
        triggered.add(103)
    if [(max(x / 200, y / 230, z / 65) < 1.015)] != [(max(x / 200, y / 230, z / 65) < 15)]:
        triggered.add(104)

    if [(abs(max(x, y / 10 * 8.7, z) - min(x, y / 10 * 8.7, z)) < 15)] != [(abs(max(x, y / 10 * 8.7, z) - min(x, y / 10 * 8.7, z)) < 125)]:
        triggered.add(105)
    if [(abs(max(x, y / 10 * 8.7, z) - min(x, y / 10 * 8.7, z)) < 15)] != [(abs(max(x, y / 10 * 8.7, z) - min(x, y / 10 * 8.7, z)) < 88)]:
        triggered.add(106)

    if [((x + y / 10 + z) > 297 and (x + y / 10 + z) < 299)] != [((x + y / 10 + z) > 297 and (x + y / 10 + z) < 99)]:
        triggered.add(107)
    if [((x + y / 10 + z) > 297 and (x + y / 10 + z) < 299)] != [((x + y / 10 + z) > 297 and (x + y / 10 + z) < 549)]:
        triggered.add(108)

    if [((x * y * z / 10000) > 31.5 and (x * y * z / 10000) < 345)] != [((x * y * z / 10000) > 31.5 and (x * y * z / 10000) < 132.5)]:
        triggered.add(109)
    if [((x * y * z / 10000) > 31.5 and (x * y * z / 10000) < 32.5)] != [((x * y * z / 10000) > 31.5 and (x * y * z / 10000) < 232.5)]:
        triggered.add(110)

    if [(abs((x + y / 10 + z) / 3 - 99.3) < 0.8)] != [(abs((x + y / 10 + z) / 3 - 99.3) < 8)]:
        triggered.add(111)
    if [(abs((x + y / 10 + z) / 3 - 99.3) < 0.8)] != [(abs((x + y / 10 + z) / 3 - 99.3) < 10.8)]:
        triggered.add(112)

    if [(x > 199 and x < 201 and y > 229 and y < 231)] != [(x > 199 and x < 201 and y > 229 and y < 2231)]:
        triggered.add(113)
    if [(x > 199 and x < 201 and y > 229 and y < 231)] != [(x > 199 and x < 201 and y > 229 and y < 1231)]:
        triggered.add(114)

    if [(z > 64.5 and z < 65.5)] != [(z > 64.5 and z < 615.5)]:
        triggered.add(115)
    if [(z > 64.5 and z < 65.5)] != [(z > 64.5 and z < 655)]:
        triggered.add(116)

    if [((x - 200) ** 2 + (y / 10 - 23) ** 2 + (z - 65) ** 2) < 9] != [((x - 200) ** 2 + (y / 10 - 23) ** 2 + (z - 65) ** 2) < 19]:
        triggered.add(117)
    if [((x - 200) ** 2 + (y / 10 - 23) ** 2 + (z - 65) ** 2) < 9] != [((x - 200) ** 2 + (y / 10 - 23) ** 2 + (z - 65) ** 2) < 91]:
        triggered.add(118)

    if [(abs(x / (y / 10) / z - 0.134) < 0.005)] != [(abs(x / (y / 10) / z - 0.134) < 10.005)]:
        triggered.add(119)
    if [(abs(x / (y / 10) / z - 0.134) < 0.005)] != [(abs(x / (y / 10) / z - 0.134) < 5)]:
        triggered.add(120)

    if [(x < 188 or y < 225 or z < 58)] != [(x < 188 or y < 225 or z < 158)]:
        triggered.add(121)
    if [(x < 188 or y < 225 or z < 58)] != [(x < 188 or y < 225 or 28 < 58)]:
        triggered.add(122)

    if [(x > 212 or y > 235 or z > 72)] != [(x > 212 or y > 235 or 12 * z > 72)]:
        triggered.add(123)
    if [(x > 212 or y > 235 or z > 72)] != [(x > 212 or y > 235 or 88 > 72)]:
        triggered.add(124)

    if [(abs(x / (y / 10) - 8.7) > 0.5)] != [(abs(x / (y / 10) - 8.7) > 10.5)]:
        triggered.add(125)
    if [(abs(x / (y / 10) - 8.7) > 0.5)] != [(abs(x / (y / 10) - 8.7) > 5)]:
        triggered.add(126)

    if [(abs(x / z - 3.08) > 0.3)] != [(abs(x / z - 3.08) > 10.3)]:
        triggered.add(127)
    if [(abs(x / z - 3.08) > 0.3)] != [(abs(x / z - 3.08) > 123)]:
        triggered.add(128)

    if [(abs((y / 10) / z - 0.354) > 0.03)] != [(abs((y / 10) / z - 0.354) > 3)]:
        triggered.add(129)
    if [(abs((y / 10) / z - 0.354) > 0.03)] != [(abs((y / 10) / z - 0.354) > 10.03)]:
        triggered.add(130)

    if [((x + y / 10 + z) < 292 or (x + y / 10 + z) > 304)] != [((x + y / 10 + z) < 292 or (x + y / 10 + z) > 3224)]:
        triggered.add(131)
    if [((x + y / 10 + z) < 292 or (x + y / 10 + z) > 304)] != [((x + y / 10 + z) < 292 or (x + y / 10 + z) > 1304)]:
        triggered.add(132)

    if [((x * y * z / 10000) < 29 or (x * y * z / 10000) > 35)] != [((x * y * z / 10000) < 29 or (x * y * z / 10000) > 135)]:
        triggered.add(133)
    if [((x * y * z / 10000) < 29 or (x * y * z / 10000) > 35)] != [((x * y * z / 10000) < 29 or (x * y * z / 10000) > 315)]:
        triggered.add(134)

    if [(x > 215)] != [(x > 1215)]:
        triggered.add(135)
    if [(x > 215)] != [(x > 2115)]:
        triggered.add(136)

    if [(x < 185)] != [(x < 1815)]:
        triggered.add(137)
    if [(x < 185)] != [(x < 1185)]:
        triggered.add(138)

    if [(y > 237)] != [(y > 2137)]:
        triggered.add(139)
    if [(y > 237)] != [(y > 2317)]:
        triggered.add(140)

    if [(y < 223)] != [(y < 2123)]:
        triggered.add(141)
    if [(y < 223)] != [(y < 2283)]:
        triggered.add(142)

    if [(z > 75)] != [(z > 751)]:
        triggered.add(143)
    if [(z > 75)] != [(z > 175)]:
        triggered.add(144)

    if [(z < 55)] != [(z < 515)]:
        triggered.add(145)
    if [(z < 55)] != [(z < 559)]:
        triggered.add(146)

    if [(x > 215 or x < 185) and (y > 237 or y < 223)] != [(x > 215 or x < 185) and (y > 237 or y < 2213)]:
        triggered.add(147)
    if [(x > 215 or x < 185) and (y > 237 or y < 223)] != [(x > 215 or x < 185) and (y > 237 or 200 < 223)]:
        triggered.add(148)

    if [(x > 215 or x < 185) and (z > 75 or z < 55)] != [(x > 215 or x < 185) and (z > 175 or z < 55)]:
        triggered.add(149)
    if [(x > 215 or x < 185) and (z > 75 or z < 55)] != [(x > 215 or x < 185) and (z > 75 or z < 551)]:
        triggered.add(150)

    if [(y > 237 or y < 223) and (z > 75 or z < 55)] != [(y > 237 or y < 223) and (z > 75 or z < 515)]:
        triggered.add(151)
    if [(y > 237 or y < 223) and (z > 75 or z < 55)] != [(y > 237 or y < 223) and (z > 175 or z < 55)]:
        triggered.add(152)

    if [(x > 215 or x < 185) and (y > 237 or y < 223) and (z > 75 or z < 55)] != [(x > 215 or x < 1185) and (y > 237 or y < 223) and (z > 75 or z < 55)]:
        triggered.add(153)
    if [(x > 215 or x < 185) and (y > 237 or y < 223) and (z > 75 or z < 55)] != [(x > 215 or x < 185) and (y > 2317 or y < 223) and (z > 75 or z < 55)]:
        triggered.add(154)

    if [(abs(x - 200) > 20) or (abs(y - 230) > 15) or (abs(z - 65) > 10)] != [(abs(x - 200) > 20) or (abs(y - 230) > 115) or (abs(z - 65) > 10)]:
        triggered.add(155)
    if [(abs(x - 200) > 20) or (abs(y - 230) > 15) or (abs(z - 65) > 10)] != [(abs(x - 200) > 20) or (abs(y - 230) > 415) or (abs(z - 65) > 10)]:
        triggered.add(156)

    if [(x > 220) or (y > 240) or (z > 80)] != [(x > 220) or (y > 240) or (z > 810)]:
        triggered.add(157)
    if [(x > 220) or (y > 240) or (z > 80)] != [(x > 220) or (y > 240) or (z > 911)]:
        triggered.add(158)
    if [(x < 175) or (y < 215) or (z < 45)] != [(x < 175) or (y < 215) or (z < 415)]:
        triggered.add(159)
    if [(x < 175) or (y < 215) or (z < 45)] != [(x < 175) or (y < 215) or (z < 145)]:
        triggered.add(160)
    return triggered


# ========== execute_Tr 指向 section9 ==========
def execute_Tr(x, y, z):
    return section9_hybrid_speed_torque_current(x, y, z)


# === Jaccard 相似度 ===
def jaccard_similarity(set1, set2):
    """Compute Jaccard similarity"""
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0


# === 路径相似度矩阵 ===
def compute_path_similarity_matrix(paths):
    """Path Similarity"""
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix


# === 目标路径组（与 section9 匹配） ===
targetPaths = [
    {1, 2, 3, 4, 5, 6, 11, 12, 14, 16, 17, 18, 19, 21, 22, 23, 25, 26, 29, 30, 31, 34, 35, 36, 37, 38, 39, 40, 41, 42,
     47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100,
     101, 102, 103, 104, 108, 112, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137,
     138, 141, 142, 143, 144, 145, 146, 147, 148, 149, 157, 158, 159, 160},
    {1, 2, 3, 4, 5, 6, 11, 12, 14, 15, 17, 18, 19, 23, 25, 26, 29, 30, 32, 34, 35, 36, 37, 38, 39, 40, 41, 42, 47, 48,
     49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102,
     103, 104, 108, 112, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 141,
     142, 143, 144, 145, 146, 147, 148, 149, 157, 158, 159, 160},
    {1, 2, 3, 4, 5, 6, 11, 12, 14, 16, 17, 18, 19, 23, 25, 26, 29, 30, 31, 34, 35, 36, 37, 38, 39, 40, 41, 42, 47, 48,
     49, 50, 53, 54, 55, 56, 57, 58, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102, 103,
     104, 108, 111, 112, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 141,
     142, 143, 144, 145, 146, 147, 148, 149, 157, 158, 159, 160},
    {1, 2, 3, 4, 5, 6, 14, 16, 17, 18, 19, 22, 25, 26, 29, 30, 31, 34, 35, 36, 37, 38, 39, 40, 41, 42, 47, 48, 49, 50,
     51, 52, 54, 55, 56, 57, 58, 63, 64, 67, 68, 69, 70, 73, 75, 76, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102,
     103, 104, 108, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 139, 140,
     141, 142, 143, 144, 145, 146, 149, 152, 154, 159, 160},
    {1, 2, 3, 4, 5, 6, 14, 16, 17, 18, 19, 21, 22, 23, 25, 26, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 47,
     48, 49, 50, 53, 54, 55, 56, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104,
     108, 111, 112, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 139, 140,
     141, 142, 143, 144, 145, 146, 149, 152, 154, 159, 160},
    {1, 2, 3, 4, 5, 6, 9, 10, 11, 12, 14, 17, 18, 19, 20, 21, 22, 25, 26, 27, 28, 29, 30, 32, 34, 35, 36, 37, 38, 39,
     40, 41, 42, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 85, 86, 87, 88, 95, 96,
     97, 98, 99, 100, 103, 104, 105, 108, 111, 112, 115, 116, 119, 120, 125, 126, 127, 128, 129, 130, 131, 132, 137,
     138, 141, 142, 143, 144, 145, 146, 157, 158, 159, 160},
    {1, 2, 3, 4, 5, 6, 11, 12, 14, 17, 18, 19, 21, 22, 23, 25, 26, 29, 30, 31, 34, 35, 36, 37, 38, 39, 40, 41, 42, 47,
     48, 49, 50, 51, 52, 54, 55, 56, 57, 58, 59, 60, 63, 64, 67, 68, 69, 70, 73, 77, 78, 79, 80, 95, 96, 97, 98, 99,
     100, 103, 104, 108, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 141,
     142, 143, 144, 145, 146, 147, 148, 149, 159, 160},
    {1, 2, 3, 4, 5, 6, 9, 10, 14, 16, 17, 18, 19, 20, 21, 22, 25, 26, 29, 30, 31, 34, 35, 36, 37, 38, 39, 40, 47, 48,
     49, 50, 51, 52, 55, 56, 57, 58, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102, 103,
     104, 105, 108, 111, 112, 115, 116, 119, 120, 121, 122, 127, 128, 129, 130, 131, 132, 137, 138, 139, 140, 141, 142,
     143, 144, 145, 146, 152, 153, 157, 158, 159, 160},
    {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 13, 16, 17, 18, 21, 22, 25, 26, 29, 30, 32, 34, 35, 36, 37, 38, 39, 40, 49, 50, 51,
     52, 55, 56, 57, 58, 63, 64, 69, 70, 77, 78, 79, 80, 87, 88, 93, 94, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104,
     105, 108, 111, 112, 115, 116, 119, 120, 121, 122, 127, 128, 129, 130, 131, 132, 137, 138, 139, 140, 141, 142, 143,
     144, 145, 146, 152, 153, 157, 158, 159, 160},
    {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 16, 17, 18, 19, 20, 21, 22, 25, 26, 29, 30, 32, 33, 34, 35, 36, 37, 38, 49, 50, 51,
     52, 53, 54, 55, 56, 57, 58, 63, 64, 69, 70, 77, 78, 79, 80, 89, 90, 95, 96, 97, 98, 99, 100, 103, 104, 105, 108,
     111, 112, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 137, 138, 139, 140, 141, 142, 143,
     144, 145, 146, 152, 153, 157, 158, 159, 160},
    {1, 2, 3, 4, 5, 6, 14, 16, 17, 18, 19, 21, 22, 23, 25, 26, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40, 47, 48, 49,
     50, 53, 54, 55, 56, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 84, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 108,
     111, 112, 115, 116, 119, 120, 121, 122, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 139, 140, 141, 142, 143,
     144, 145, 146, 149, 152, 154, 159, 160},
    {1, 2, 3, 4, 5, 6, 11, 12, 14, 16, 17, 18, 19, 21, 22, 23, 24, 25, 26, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40,
     41, 42, 47, 48, 49, 50, 54, 55, 56, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104,
     108, 111, 112, 115, 116, 119, 120, 121, 122, 125, 126, 129, 130, 131, 132, 135, 136, 137, 138, 141, 142, 143, 144,
     145, 146, 147, 148, 149, 159, 160},
    {1, 2, 3, 4, 5, 6, 14, 16, 17, 18, 19, 21, 22, 23, 24, 25, 26, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42,
     43, 44, 47, 48, 49, 50, 54, 55, 56, 67, 68, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102, 103,
     104, 108, 112, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 131, 132, 135, 136, 137, 138, 139, 140, 141, 142,
     145, 146, 150, 151, 159, 160},
    {1, 2, 3, 4, 9, 10, 11, 12, 13, 15, 17, 18, 25, 26, 27, 28, 29, 30, 32, 34, 35, 36, 37, 38, 39, 40, 49, 50, 51, 52,
     55, 56, 57, 58, 63, 64, 69, 70, 77, 78, 79, 80, 81, 82, 85, 86, 87, 88, 95, 96, 97, 98, 99, 100, 101, 102, 103,
     104, 105, 108, 111, 112, 115, 116, 119, 120, 121, 122, 127, 128, 129, 130, 131, 132, 137, 138, 141, 142, 143, 144,
     145, 146, 157, 158, 159, 160},
    {1, 2, 3, 4, 5, 6, 14, 16, 17, 18, 19, 22, 23, 24, 25, 26, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40, 47, 48, 49,
     50, 53, 54, 55, 56, 69, 70, 73, 74, 77, 78, 79, 80, 83, 84, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 108, 111,
     112, 115, 116, 119, 120, 121, 122, 129, 130, 131, 132, 135, 136, 137, 138, 139, 140, 141, 142, 143, 144, 145, 146,
     149, 152, 154, 159, 160},
    {1, 2, 3, 4, 5, 6, 9, 10, 11, 12, 13, 15, 17, 18, 23, 25, 26, 27, 28, 32, 33, 34, 35, 36, 37, 38, 49, 50, 55, 56,
     57, 58, 63, 64, 69, 70, 77, 78, 79, 80, 81, 82, 85, 86, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 105, 108, 111,
     112, 113, 114, 115, 116, 119, 120, 121, 122, 127, 128, 129, 130, 131, 132, 137, 138, 141, 142, 143, 144, 145, 146,
     157, 158, 159, 160},
    {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 13, 16, 17, 18, 22, 23, 24, 25, 26, 32, 33, 34, 35, 36, 49, 50, 55, 56, 63, 64, 69,
     70, 77, 78, 79, 80, 87, 88, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 108, 111, 112, 113, 114,
     115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 137, 138, 139, 140, 141, 142, 143, 144, 145, 146, 152,
     153, 159, 160},
    {1, 2, 3, 4, 5, 6, 14, 16, 19, 22, 23, 24, 25, 26, 29, 30, 31, 33, 34, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47,
     48, 49, 50, 54, 55, 56, 67, 68, 69, 70, 77, 78, 79, 80, 97, 98, 99, 100, 103, 104, 108, 111, 112, 119, 120, 121,
     122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 139, 140, 141, 142, 145, 146, 150, 151, 159, 160},
    {1, 2, 5, 6, 14, 16, 17, 18, 19, 22, 23, 24, 25, 26, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40, 47, 48, 49, 50, 53,
     54, 69, 70, 73, 74, 77, 78, 79, 80, 83, 84, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 108, 111, 112, 115, 116,
     119, 120, 121, 122, 131, 132, 135, 136, 137, 138, 139, 140, 141, 142, 145, 146, 150, 151, 155, 156, 159, 160},
    {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 13, 16, 17, 18, 22, 23, 24, 25, 26, 32, 33, 34, 35, 36, 53, 54, 55, 56, 63, 64, 69,
     70, 77, 78, 79, 80, 87, 88, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 107, 113, 114, 115, 116,
     118, 119, 120, 121, 122, 125, 126, 127, 128, 137, 138, 139, 140, 141, 142, 145, 146, 151, 155, 156, 159, 160},
    {1, 2, 5, 6, 9, 10, 14, 17, 18, 19, 20, 22, 23, 27, 28, 29, 30, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 49, 50, 54,
     55, 56, 57, 58, 63, 64, 69, 70, 73, 74, 79, 80, 87, 88, 95, 96, 97, 98, 99, 100, 103, 104, 108, 109, 111, 112, 115,
     116, 119, 120, 125, 126, 127, 128, 129, 130, 131, 132, 137, 138, 143, 144, 145, 146, 152, 153, 157, 158},
    {1, 2, 3, 4, 5, 6, 14, 16, 19, 22, 23, 24, 25, 26, 29, 30, 31, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49,
     50, 54, 55, 56, 57, 58, 61, 62, 65, 67, 68, 69, 70, 97, 98, 99, 100, 103, 104, 108, 109, 111, 112, 119, 120, 125,
     126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 139, 140, 141, 142, 145, 146, 150, 151, 159, 160},
    {1, 2, 3, 4, 5, 6, 14, 16, 19, 21, 22, 23, 24, 25, 26, 29, 30, 31, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48,
     49, 50, 54, 55, 56, 65, 69, 70, 97, 98, 99, 100, 103, 104, 108, 109, 111, 112, 119, 120, 121, 122, 125, 126, 127,
     128, 129, 130, 131, 132, 134, 135, 136, 137, 138, 141, 142, 145, 146, 147, 148, 150, 159, 160},
    {1, 2, 3, 4, 5, 6, 7, 8, 13, 16, 17, 18, 22, 23, 24, 25, 26, 32, 35, 36, 53, 54, 69, 70, 79, 80, 87, 88, 91, 92, 93,
     94, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 109, 111, 112, 115, 116, 117, 118, 119, 120, 121, 122, 125, 126,
     137, 138, 139, 140, 141, 142, 145, 146, 151, 155, 156, 159, 160},
    {1, 2, 3, 4, 5, 6, 9, 10, 11, 12, 14, 15, 17, 18, 19, 20, 23, 24, 25, 26, 32, 33, 34, 35, 36, 39, 40, 49, 50, 53,
     54, 69, 70, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 108, 109, 111, 112, 115, 116, 119, 120, 121, 122,
     123, 124, 129, 130, 137, 138, 141, 142, 145, 146, 159, 160},
    {3, 4, 5, 6, 14, 16, 19, 21, 22, 31, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 54, 56, 57, 58, 61, 62, 65, 66,
     67, 68, 69, 70, 79, 80, 99, 100, 103, 104, 108, 109, 110, 111, 112, 119, 120, 125, 126, 127, 128, 129, 130, 131,
     132, 134, 135, 136, 137, 138, 141, 142, 147, 148},
    {1, 2, 5, 6, 17, 18, 19, 22, 27, 28, 33, 34, 35, 36, 51, 52, 54, 56, 57, 58, 63, 64, 65, 69, 70, 71, 72, 89, 90, 95,
     96, 97, 98, 99, 100, 103, 104, 105, 106, 109, 112, 115, 116, 119, 120, 125, 126, 127, 128, 129, 130, 134, 143, 144,
     145, 146, 149, 152, 157, 158},
    {1, 2, 3, 4, 14, 19, 21, 22, 31, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 54, 57, 58, 59, 60, 61, 62, 65, 66, 75, 76,
     77, 78, 79, 80, 99, 100, 103, 104, 110, 111, 112, 119, 120, 125, 126, 127, 128, 129, 130, 133, 134, 135, 136, 137,
     138, 141, 142, 147, 148},
]


# === 路径分组 ===
def group_paths_by_similarity(paths):
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.mean(avg_sim_scores)

    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]
    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)

    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group


# === 优先级经验回放 ===
class PrioritizedExperienceReplay:
    def __init__(self, capacity=10000, alpha=0.6, beta_start=0.4, beta_frames=100000):
        self.capacity = capacity
        self.alpha = alpha
        self.beta_start = beta_start
        self.beta_frames = beta_frames
        self.frame = 1

        self.buffer = []
        self.priorities = np.zeros(capacity, dtype=np.float32)
        self.pos = 0
        self.size = 0

        self.max_priority = 1.0
        self.min_priority = 1.0

    def beta(self):
        """beta()"""
        return min(1.0, self.beta_start + (1.0 - self.beta_start) * self.frame / self.beta_frames)

    def append(self, experience):
        """添加经验，设置默认优先级"""
        if len(self.buffer) < self.capacity:
            self.buffer.append(experience)
        else:
            self.buffer[self.pos] = experience

        self.priorities[self.pos] = self.max_priority

        self.pos = (self.pos + 1) % self.capacity
        self.size = min(self.size + 1, self.capacity)

    def sample(self, batch_size):
        """采样，返回去重的样本"""
        if self.size < batch_size:
            return [], [], []

        priorities = self.priorities[:self.size]
        probs = priorities ** self.alpha
        probs /= probs.sum()

        indices = np.random.choice(self.size, batch_size, p=probs, replace=False)

        # 去重
        unique_batch = []
        unique_indices = []
        seen_states = set()

        for idx in indices:
            experience = self.buffer[idx]
            state_tensor = experience[0]
            state_tuple = tuple(state_tensor.cpu().numpy().flatten())

            if state_tuple not in seen_states:
                seen_states.add(state_tuple)
                unique_batch.append(experience)
                unique_indices.append(idx)

        if len(unique_batch) < batch_size:
            remaining_indices = [i for i in range(self.size) if i not in unique_indices]
            if remaining_indices:
                remaining_probs = priorities[remaining_indices] ** self.alpha
                remaining_probs /= remaining_probs.sum()

                needed = batch_size - len(unique_batch)
                additional_indices = np.random.choice(
                    remaining_indices,
                    min(needed, len(remaining_indices)),
                    p=remaining_probs,
                    replace=False
                )

                for idx in additional_indices:
                    experience = self.buffer[idx]
                    state_tensor = experience[0]
                    state_tuple = tuple(state_tensor.cpu().numpy().flatten())

                    if state_tuple not in seen_states:
                        seen_states.add(state_tuple)
                        unique_batch.append(experience)
                        unique_indices.append(idx)

        total = len(self.buffer)
        unique_indices = np.array(unique_indices)
        weights = (total * probs[unique_indices]) ** (-self.beta())
        weights /= weights.max()

        self.frame += 1

        return unique_batch, unique_indices, weights

    def update_priorities(self, indices, priorities):
        """更新优先级"""
        for idx, priority in zip(indices, priorities):
            if idx < self.size:
                self.priorities[idx] = priority
                self.max_priority = max(self.max_priority, priority)
                self.min_priority = min(self.min_priority, priority)

    def get_priority_statistics(self):
        """获取优先级统计信息"""
        if self.size == 0:
            return None

        priorities = self.priorities[:self.size]
        mean_priority = np.mean(priorities)
        max_priority = np.max(priorities)
        min_priority = np.min(priorities)

        high_priority_ratio = np.mean(priorities > mean_priority)

        return {
            'mean_priority': mean_priority,
            'max_priority': max_priority,
            'min_priority': min_priority,
            'high_priority_ratio': high_priority_ratio
        }

    def __len__(self):
        return self.size

    def get_high_reward_samples(self, target_path, num_samples=20):
        """获取高奖励样本"""
        if len(self.buffer) == 0:
            return []

        samples_with_recalculated_scores = []
        for experience in self.buffer:
            state_tensor = experience[0]
            state_tuple = tuple(state_tensor.cpu().numpy().flatten())
            triggered = execute_Tr(*state_tuple)
            new_reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_recalculated_scores.append((state_tuple, new_reward, sim, triggered))

        samples_with_recalculated_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_recalculated_scores[:num_samples]


def load_path_data(file_path):
    """从文件读取状态（浮点数）"""
    path_data = []
    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines[2:]:
            parts = line.strip().split("\t")
            state = tuple(map(float, parts[0].split()))
            path_data.append(state)
    return path_data


# === DQN 网络 ===
class DQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, 128)
        self.fc2 = nn.Linear(128, 64)
        self.fc3 = nn.Linear(64, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)


# === 优先级 DQN Agent ===
class PrioritizedDQNAgent:
    def __init__(self, state_dim, action_dim, replay_buffer, gamma=0.99, epsilon=1.0,
                 epsilon_decay=0.995, epsilon_min=0.1, learning_rate=0.001):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        """解码动作索引为(dx, dy, dz)的步长（整数）"""
        delta_values = [35, 25, 10, 5, 2, -2, -5, -10, -25, -35]
        dim = action_idx // 10
        delta_idx = action_idx % 10

        if dim == 0:
            return (delta_values[delta_idx], 0, 0)
        elif dim == 1:
            return (0, delta_values[delta_idx], 0)
        elif dim == 2:
            return (0, 0, delta_values[delta_idx])
        else:
            return (0, 0, 0)

    def act(self, state):
        if random.random() < self.epsilon:
            return random.randrange(self.action_dim)
        normalized_state = normalize_state(state)
        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)
        return torch.argmax(q_values, dim=1).item()

    def store_transition(self, state, action, reward, next_state, done):
        """存储经验，计算TD误差作为优先级"""
        normalized_state = normalize_state(state)
        normalized_next_state = normalize_state(next_state)

        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        next_state_tensor = torch.tensor(normalized_next_state, dtype=torch.float32).unsqueeze(0).to(device)

        with torch.no_grad():
            q_values = self.model(state_tensor)
            next_q_values = self.target_model(next_state_tensor)
            max_next_q_values = next_q_values.max(1)[0]
            target_q_values = reward + (self.gamma * max_next_q_values * (1 - done))
            td_error = torch.abs(q_values[0][action] - target_q_values).item()

        self.replay_buffer.append((state_tensor, action, reward, next_state_tensor, done, td_error))
        return td_error

    def train(self, batch_size=32):
        """训练网络"""
        if len(self.replay_buffer) < batch_size:
            return

        batch, batch_indices, is_weights = self.replay_buffer.sample(batch_size)
        if not batch:
            return

        states, actions, rewards, next_states, dones, _ = zip(*batch)

        states = torch.tensor(np.array([s.cpu().numpy().flatten() for s in states]), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array([ns.cpu().numpy().flatten() for ns in next_states]),
                                   dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)
        is_weights = torch.tensor(is_weights, dtype=torch.float32).to(device)

        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))

        td_errors = torch.abs(current_q_values - target_q_values)
        loss = (is_weights * (current_q_values - target_q_values) ** 2).mean()

        self.optimizer.zero_grad()
        loss.backward()
        self.optimizer.step()

        new_priorities = td_errors.detach().cpu().numpy() + 1e-6
        self.replay_buffer.update_priorities(batch_indices, new_priorities)

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())


# === 生成相似路径样本（浮点数状态） ===
def generate_samples_for_similar_paths(similar_group_indices, num_total=2000, top_k=200):
    def jaccard_similarity_local(a, b):
        if not a and not b:
            return 1.0
        return len(a & b) / len(a | b) if a | b else 0.0

    def compute_robustness(state, path):
        base = execute_Tr(*state)
        if not base:
            return 0.0
        rob, neighbors = 0.0, 0
        for ddx in [-1, 0, 1]:
            for ddy in [-1, 0, 1]:
                for ddz in [-1, 0, 1]:
                    if ddx == ddy == ddz == 0:
                        continue
                    neighbor_state = (state[0] + ddx, state[1] + ddy, state[2] + ddz)
                    if not is_valid_state(neighbor_state):
                        continue
                    neighbor = clip_state(neighbor_state)
                    n_trig = execute_Tr(*neighbor)
                    if not n_trig:
                        continue
                    rob += jaccard_similarity_local(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"prioritized_path{path_id}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Prioritized Path {path_id}\n")
            f.write("dx dy dz\tScore\tSimilarity\tLengthDiff\tRobustness\n")
            for s in samples:
                dx, dy, dz = s[0]
                f.write(f"{dx:.2f} {dy:.2f} {dz:.2f}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_prioritized"
    for path_idx in similar_group_indices:
        path = targetPaths[path_idx]
        samples = []
        attempts = 0
        while len(samples) < top_k and attempts < num_total * 5:
            attempts += 1
            state = (
                random.uniform(X_MIN, X_MAX),
                random.uniform(Y_MIN, Y_MAX),
                random.uniform(Z_MIN, Z_MAX)
            )
            triggered = execute_Tr(*state)
            if not triggered:
                continue
            sim = jaccard_similarity_local(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            score = 0.55 * sim + 0.25 * len_diff + 0.2 * rob
            samples.append((state, score, sim, len_diff, rob))
        if samples:
            samples.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples[:top_k], base_dir=base_dir)


# === 训练相似路径组 ===
def prioritized_generate_and_train_for_similar_paths(agent, similar_group, path_documents, episodes=500,
                                                     batch_size=32, steps_per_test=3, replay_times=1,
                                                     is_isolated=False):
    trained_paths = set()
    global_replay_count = 0

    for episode in range(episodes):
        episode_reward = 0
        episode_similarities = []
        episode_td_errors = []

        for path_idx in similar_group:
            if path_idx in trained_paths:
                continue

            file_path = os.path.join(path_documents,
                                     f"prioritized_path{path_idx + 1}{'_isolated' if is_isolated else ''}.txt")
            path_data = load_path_data(file_path)
            target_path = targetPaths[path_idx]

            BATCH_SIZE = 50
            N_SAMPLES = 200
            N_STEPS = 3
            N_BATCHES = 4
            PATH_REPEAT = 5

            for repeat in range(PATH_REPEAT):
                for batch_idx in range(N_BATCHES):
                    batch_start = batch_idx * BATCH_SIZE
                    batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)

                    for test_data_idx in range(batch_start, batch_end):
                        if test_data_idx >= len(path_data):
                            break

                        state = path_data[test_data_idx]
                        prev_state = None
                        prev_triggered = None
                        prev_reward = None

                        for step in range(N_STEPS):
                            legal_actions = []
                            for a in range(agent.action_dim):
                                ddx, ddy, ddz = agent.decode_action(a)
                                cand_next = (state[0] + ddx, state[1] + ddy, state[2] + ddz)
                                if is_valid_state(cand_next):
                                    legal_actions.append(a)
                            if not legal_actions:
                                break

                            if random.random() < agent.epsilon:
                                action = random.choice(legal_actions)
                            else:
                                normalized_state = normalize_state(state)
                                state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(
                                    device)
                                with torch.no_grad():
                                    q_values = agent.model(state_tensor)[0]
                                action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                            ddx, ddy, ddz = agent.decode_action(action)
                            next_state = clip_state((state[0] + ddx, state[1] + ddy, state[2] + ddz))

                            triggered = execute_Tr(*next_state)
                            reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                            done = (step == N_STEPS - 1)

                            td_error = agent.store_transition(state, action, reward, next_state, done)

                            priority_stats = agent.replay_buffer.get_priority_statistics()
                            current_priority = priority_stats['mean_priority'] if priority_stats else 0

                            prioritized_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                    current_priority, None)

                            episode_similarities.append(jaccard_similarity(triggered, target_path))
                            episode_td_errors.append(td_error)

                            if prev_reward is not None:
                                prioritized_metrics.record_action_improvement(reward, prev_reward)

                            prev_state = state
                            prev_triggered = triggered
                            prev_reward = reward
                            state = next_state
                            episode_reward += reward

                    if len(agent.replay_buffer) >= batch_size:
                        agent.train(batch_size)
                        global_replay_count += 1

                        if global_replay_count % 2 == 0:
                            agent.update_target_model()

            trained_paths.add(path_idx)

        avg_similarity = np.mean(episode_similarities) if episode_similarities else 0
        avg_td_error = np.mean(episode_td_errors) if episode_td_errors else 0
        priority_stats = agent.replay_buffer.get_priority_statistics()

        prioritized_metrics.record_episode_metrics(episode, episode_reward, avg_similarity, avg_td_error,
                                                   agent.epsilon, "similar", priority_stats)

        if len(trained_paths) == len(similar_group):
            break

    return agent


# === 生成独立路径样本（浮点数） ===
def generate_samples_for_isolated_paths_prioritized(agent_similar, isolated_group_indices, num_total=2000, top_k=200):
    def compute_q_value_normalized_complement(state, agent):
        """计算Q值归一化后的补值"""
        normalized_state = normalize_state(state)
        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = agent.model(state_tensor)

        all_q_values = q_values[0].cpu().numpy()

        q_min = all_q_values.min()
        q_max = all_q_values.max()

        if q_max - q_min > 1e-6:
            normalized_q = (all_q_values.max() - q_min) / (q_max - q_min)
        else:
            normalized_q = 0.0

        complement_q = 1.0 - normalized_q

        return complement_q

    def compute_robustness(state, path):
        base = execute_Tr(*state)
        if not base:
            return 0.0
        rob, neighbors = 0.0, 0
        for ddx in [-1, 0, 1]:
            for ddy in [-1, 0, 1]:
                for ddz in [-1, 0, 1]:
                    if ddx == ddy == ddz == 0:
                        continue
                    neighbor_state = (state[0] + ddx, state[1] + ddy, state[2] + ddz)
                    if not is_valid_state(neighbor_state):
                        continue
                    neighbor = clip_state(neighbor_state)
                    n_trig = execute_Tr(*neighbor)
                    if not n_trig:
                        continue
                    rob += jaccard_similarity(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"prioritized_path{path_id}_isolated.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Prioritized Isolated Path {path_id}\n")
            f.write("dx dy dz\tScore\tSimilarity\tLengthDiff\tRobustness\tQ_value_complement\n")
            for s in samples:
                dx, dy, dz = s[0]
                f.write(f"{dx:.2f} {dy:.2f} {dz:.2f}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\t{s[5]:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_prioritized"
    for path_idx in isolated_group_indices:
        path = targetPaths[path_idx]
        samples = []
        attempts = 0
        while len(samples) < top_k and attempts < num_total * 5:
            attempts += 1
            state = (
                random.uniform(X_MIN, X_MAX),
                random.uniform(Y_MIN, Y_MAX),
                random.uniform(Z_MIN, Z_MAX)
            )
            triggered = execute_Tr(*state)
            if not triggered:
                continue
            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            q_value_complement = compute_q_value_normalized_complement(state, agent_similar)
            score = 0.28 * sim + 0.1 * len_diff + 0.19 * rob + 0.43 * q_value_complement
            samples.append((state, score, sim, len_diff, rob, q_value_complement))
        if samples:
            samples.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples[:top_k], base_dir=base_dir)


# === 训练独立路径组 ===
def prioritized_generate_and_train_for_isolated_paths(agent_similar, agent_isolated, similar_group,
                                                      isolated_group, path_documents, episodes=500, batch_size=32,
                                                      is_isolated=True):
    trained_paths = set()
    global_replay_count = 0

    stage1_samples_pool = {}
    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_similar.replay_buffer.get_high_reward_samples(target_path, num_samples=100)
        stage1_samples_pool[path_idx] = high_reward_samples

    for episode in range(episodes):
        episode_reward = 0
        episode_similarities = []
        episode_td_errors = []

        for path_idx in isolated_group:
            if path_idx in trained_paths:
                continue

            file_path = os.path.join(path_documents, f"prioritized_path{path_idx + 1}_isolated.txt")
            stage2_path_data = load_path_data(file_path)
            stage1_samples = stage1_samples_pool.get(path_idx, [])
            target_path = targetPaths[path_idx]

            BATCH_SIZE = 50
            N_SAMPLES_STAGE2 = min(140, len(stage2_path_data))
            N_SAMPLES_STAGE1 = min(60, len(stage1_samples))
            N_STEPS = 3
            N_BATCHES_STAGE2 = (N_SAMPLES_STAGE2 + BATCH_SIZE - 1) // BATCH_SIZE
            N_BATCHES_STAGE1 = (N_SAMPLES_STAGE1 + BATCH_SIZE - 1) // BATCH_SIZE
            PATH_REPEAT = 5

            for repeat in range(PATH_REPEAT):
                # Stage2
                for batch_idx in range(N_BATCHES_STAGE2):
                    batch_start = batch_idx * BATCH_SIZE
                    batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES_STAGE2)

                    for test_data_idx in range(batch_start, batch_end):
                        if test_data_idx >= len(stage2_path_data):
                            break

                        state = stage2_path_data[test_data_idx]
                        prev_state = None
                        prev_triggered = None
                        prev_reward = None

                        for step in range(N_STEPS):
                            legal_actions = []
                            for a in range(agent_isolated.action_dim):
                                ddx, ddy, ddz = agent_isolated.decode_action(a)
                                cand_next = (state[0] + ddx, state[1] + ddy, state[2] + ddz)
                                if is_valid_state(cand_next):
                                    legal_actions.append(a)

                            if not legal_actions:
                                break

                            if random.random() < agent_isolated.epsilon:
                                action = random.choice(legal_actions)
                            else:
                                normalized_state = normalize_state(state)
                                state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(
                                    device)
                                with torch.no_grad():
                                    q_values = agent_isolated.model(state_tensor)[0]
                                action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                            ddx, ddy, ddz = agent_isolated.decode_action(action)
                            next_state = clip_state((state[0] + ddx, state[1] + ddy, state[2] + ddz))

                            triggered = execute_Tr(*next_state)
                            reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)

                            if target_path.issubset(triggered):
                                reward += 2.0

                            done = (step == N_STEPS - 1)

                            td_error = agent_isolated.store_transition(state, action, reward, next_state, done)

                            priority_stats = agent_isolated.replay_buffer.get_priority_statistics()
                            current_priority = priority_stats['mean_priority'] if priority_stats else 0

                            prioritized_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                    current_priority, None)

                            episode_similarities.append(jaccard_similarity(triggered, target_path))
                            episode_td_errors.append(td_error)

                            if prev_reward is not None:
                                prioritized_metrics.record_action_improvement(reward, prev_reward)

                            prev_state = state
                            prev_triggered = triggered
                            prev_reward = reward
                            state = next_state
                            episode_reward += reward

                    if len(agent_isolated.replay_buffer) >= batch_size:
                        agent_isolated.train(batch_size)
                        global_replay_count += 1

                        if global_replay_count % 2 == 0:
                            agent_isolated.update_target_model()

                # Stage1
                if stage1_samples:
                    for batch_idx in range(N_BATCHES_STAGE1):
                        batch_start = batch_idx * BATCH_SIZE
                        batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES_STAGE1)

                        for sample_idx in range(batch_start, batch_end):
                            if sample_idx >= len(stage1_samples):
                                break

                            stage1_state_tuple, _, _, _ = stage1_samples[sample_idx]
                            state = stage1_state_tuple
                            prev_state = None
                            prev_triggered = None
                            prev_reward = None

                            for step in range(N_STEPS):
                                legal_actions = []
                                for a in range(agent_isolated.action_dim):
                                    ddx, ddy, ddz = agent_isolated.decode_action(a)
                                    cand_next = (state[0] + ddx, state[1] + ddy, state[2] + ddz)
                                    if is_valid_state(cand_next):
                                        legal_actions.append(a)

                                if not legal_actions:
                                    break

                                if random.random() < agent_isolated.epsilon:
                                    action = random.choice(legal_actions)
                                else:
                                    normalized_state = normalize_state(state)
                                    state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(
                                        device)
                                    with torch.no_grad():
                                        q_values = agent_isolated.model(state_tensor)[0]
                                    action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                                ddx, ddy, ddz = agent_isolated.decode_action(action)
                                next_state = clip_state((state[0] + ddx, state[1] + ddy, state[2] + ddz))

                                triggered = execute_Tr(*next_state)
                                reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                                reward *= 0.8

                                done = (step == N_STEPS - 1)

                                td_error = agent_isolated.store_transition(state, action, reward, next_state, done)

                                priority_stats = agent_isolated.replay_buffer.get_priority_statistics()
                                current_priority = priority_stats['mean_priority'] if priority_stats else 0

                                prioritized_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                        current_priority, None)

                                episode_similarities.append(jaccard_similarity(triggered, target_path))
                                episode_td_errors.append(td_error)

                                if prev_reward is not None:
                                    prioritized_metrics.record_action_improvement(reward, prev_reward)

                                prev_state = state
                                prev_triggered = triggered
                                prev_reward = reward
                                state = next_state
                                episode_reward += reward

                        if len(agent_isolated.replay_buffer) >= batch_size:
                            agent_isolated.train(batch_size)
                            global_replay_count += 1

                            if global_replay_count % 2 == 0:
                                agent_isolated.update_target_model()

            trained_paths.add(path_idx)

        avg_similarity = np.mean(episode_similarities) if episode_similarities else 0
        avg_td_error = np.mean(episode_td_errors) if episode_td_errors else 0
        priority_stats = agent_isolated.replay_buffer.get_priority_statistics()

        prioritized_metrics.record_episode_metrics(episode, episode_reward, avg_similarity, avg_td_error,
                                                   agent_isolated.epsilon, "isolated", priority_stats)

        if len(trained_paths) == len(isolated_group):
            break

    return agent_isolated


# === 指标写入 Excel ===
def append_performance_metrics_to_excel(metrics_collector, filepath, run_number):
    """MetricExcel"""
    avg_td_error = np.mean(metrics_collector.td_errors) if metrics_collector.td_errors else 0
    total_reward = metrics_collector.total_reward
    action_improvement_rate = np.mean(
        metrics_collector.action_improvements) if metrics_collector.action_improvements else 0
    avg_final_similarity = np.mean(
        metrics_collector.final_output_similarities) if metrics_collector.final_output_similarities else 0
    avg_episode_reward = np.mean(metrics_collector.episode_rewards) if metrics_collector.episode_rewards else 0
    reward_std = np.std(metrics_collector.episode_rewards) if metrics_collector.episode_rewards else 0
    training_time = metrics_collector.end_time - metrics_collector.start_time if metrics_collector.end_time and metrics_collector.start_time else 0
    avg_memory_usage = np.mean(metrics_collector.episode_memory_usage) if metrics_collector.episode_memory_usage else 0
    per_step_time = training_time / metrics_collector.step_count * 1000 if metrics_collector.step_count > 0 else 0
    avg_priority = np.mean(metrics_collector.priority_statistics) if metrics_collector.priority_statistics else 0
    avg_importance_weight = np.mean(metrics_collector.importance_weights) if metrics_collector.importance_weights else 0

    new_row = {
        'Run': f"Run {run_number}",
        'final samplesAverage Similarity': f"{avg_final_similarity:.4f}",
        '': f"{total_reward:,.2f}",
        'episode': f"{avg_episode_reward:,.4f}",
        'Standard deviation': f"{reward_std:,.4f}",
        'TD': f"{avg_td_error:.4f}",
        '': f"{action_improvement_rate:.4f}",
        '(%)': f"{action_improvement_rate * 100:.2f}%",
        '': f"{metrics_collector.step_count:,}",
        'final samples': f"{len(metrics_collector.final_output_similarities)}",
        'Training Total Time( seconds)': f"{training_time:.2f}",
        'Training Total Time( minutes)': f"{training_time / 60:.2f}",
        '(MB)': f"{avg_memory_usage:.2f}",
        '(ms)': f"{per_step_time:.2f}",
        '': f"{avg_priority:.4f}",
        '': f"{avg_importance_weight:.4f}"
    }

    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    if os.path.exists(filepath):
        df = pd.read_excel(filepath, sheet_name='Metric')
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    else:
        df = pd.DataFrame([new_row])

    with pd.ExcelWriter(filepath, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, sheet_name='Metric', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Metric']

        worksheet.column_dimensions['A'].width = 15
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
            worksheet.column_dimensions[col].width = 18

        header_font = Font(bold=True, size=11)
        for cell in worksheet[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    print(f"Run {run_number} runMetric: {filepath}")


def append_final_samples_to_excel(agent_similar, agent_isolated, similar_group, isolated_group, targetPaths, filepath,
                                  run_number):
    """Excel"""
    new_samples = []

    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_similar.replay_buffer.get_high_reward_samples(target_path, num_samples=20)

        for state_tuple, reward, sim, triggered in high_reward_samples:
            new_samples.append({
                'Run': f"Run {run_number}",
                'Path ': 'Similar path group',
                'Path ID': path_idx + 1,
                'dx': state_tuple[0],
                'dy': state_tuple[1],
                'dz': state_tuple[2],
                'Similarity': f"{sim:.4f}",
                '': f"{reward:.2f}",
                '': len(triggered),
                '': len(target_path),
                '': str(sorted(triggered)),
                '': str(sorted(target_path))
            })

    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)

        for state_tuple, reward, sim, triggered in high_reward_samples:
            new_samples.append({
                'Run': f"Run {run_number}",
                'Path ': 'Isolated path group',
                'Path ID': path_idx + 1,
                'dx': state_tuple[0],
                'dy': state_tuple[1],
                'dz': state_tuple[2],
                'Similarity': f"{sim:.4f}",
                '': f"{reward:.2f}",
                '': len(triggered),
                '': len(target_path),
                '': str(sorted(triggered)),
                '': str(sorted(target_path))
            })

    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    if os.path.exists(filepath):
        df = pd.read_excel(filepath, sheet_name='final samples')
        df = pd.concat([df, pd.DataFrame(new_samples)], ignore_index=True)
    else:
        df = pd.DataFrame(new_samples)

    with pd.ExcelWriter(filepath, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, sheet_name='final samples', index=False)
        workbook = writer.book
        worksheet = writer.sheets['final samples']

        column_widths = {
            'A': 12, 'B': 15, 'C': 12, 'D': 10, 'E': 10, 'F': 10,
            'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 40, 'L': 40
        }
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        header_font = Font(bold=True, size=11)
        for cell in worksheet[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    print(f"Run {run_number} run: {filepath}")


# === 单次实验运行 ===
def run_single_experiment(run_number, results_save_dir):
    print(f"\n{'=' * 80}")
    print(f"Start run  {run_number}  run(DQN)")
    print(f"{'=' * 80}\n")

    prioritized_metrics.reset()
    prioritized_metrics.start_training()

    model_path_similar = os.path.join(results_save_dir, f"run{run_number}_similar.pth")
    model_path_isolated = os.path.join(results_save_dir, f"run{run_number}_isolated.pth")
    path_documents = r"D:\Experiment\CNN\DQNNEW\path_samples_prioritized"

    similar_group, isolated_group = group_paths_by_similarity(targetPaths)

    if run_number == 1:
        generate_samples_for_similar_paths(similar_group, num_total=2000, top_k=200)

    replay_buffer = PrioritizedExperienceReplay(capacity=10000, alpha=0.6, beta_start=0.4, beta_frames=100000)
    state_dim = 3
    action_dim = 30
    agent = PrioritizedDQNAgent(state_dim, action_dim, replay_buffer)

    agent = prioritized_generate_and_train_for_similar_paths(agent, similar_group, path_documents, episodes=500,
                                                             batch_size=32, is_isolated=False)

    os.makedirs(os.path.dirname(model_path_similar), exist_ok=True)
    torch.save({
        'model_state_dict': agent.model.state_dict(),
        'optimizer_state_dict': agent.optimizer.state_dict(),
        'epsilon': agent.epsilon
    }, model_path_similar)

    if run_number == 1:
        generate_samples_for_isolated_paths_prioritized(agent, isolated_group, num_total=2000, top_k=200)

    isolated_replay_buffer = PrioritizedExperienceReplay(capacity=15000, alpha=0.6, beta_start=0.4, beta_frames=100000)
    agent_isolated = PrioritizedDQNAgent(state_dim, action_dim, isolated_replay_buffer)

    try:
        checkpoint = torch.load(model_path_similar)
        agent_isolated.model.load_state_dict(checkpoint['model_state_dict'])
        agent_isolated.target_model.load_state_dict(checkpoint['model_state_dict'])
        agent_isolated.epsilon = checkpoint.get('epsilon', 0.5)
    except Exception as e:
        pass

    agent_isolated = prioritized_generate_and_train_for_isolated_paths(
        agent_similar=agent,
        agent_isolated=agent_isolated,
        similar_group=similar_group,
        isolated_group=isolated_group,
        path_documents=path_documents,
        episodes=500,
        batch_size=32,
        is_isolated=True
    )

    os.makedirs(os.path.dirname(model_path_isolated), exist_ok=True)
    torch.save({
        'model_state_dict': agent_isolated.model.state_dict(),
        'optimizer_state_dict': agent_isolated.optimizer.state_dict(),
        'epsilon': agent_isolated.epsilon
    }, model_path_isolated)

    prioritized_metrics.end_training()

    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, _, sim, triggered in high_reward_samples:
            prioritized_metrics.record_final_output_sample(triggered, target_path)

    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, _, sim, triggered in high_reward_samples:
            prioritized_metrics.record_final_output_sample(triggered, target_path)

    performance_excel_path = os.path.join(results_save_dir, "Metric_.xlsx")
    samples_excel_path = os.path.join(results_save_dir, "final samples_.xlsx")

    append_performance_metrics_to_excel(prioritized_metrics, performance_excel_path, run_number)
    append_final_samples_to_excel(agent, agent_isolated, similar_group, isolated_group, targetPaths, samples_excel_path,
                                  run_number)

    avg_similarity = np.mean(
        prioritized_metrics.final_output_similarities) if prioritized_metrics.final_output_similarities else 0
    training_time = prioritized_metrics.end_time - prioritized_metrics.start_time
    avg_priority = np.mean(prioritized_metrics.priority_statistics) if prioritized_metrics.priority_statistics else 0
    print(f"\nRun  {run_number}  runcompleted:")
    print(f"  Average Similarity: {avg_similarity:.4f}")
    print(f"  Training Time: {training_time:.2f} seconds")
    print(f"  : {prioritized_metrics.step_count}")
    print(f"  : {avg_priority:.4f}")


if __name__ == "__main__":
    results_save_dir = r"D:\Experiment\CNN\DQNNEW\results\prioritized_results"
    os.makedirs(results_save_dir, exist_ok=True)

    NUM_RUNS = 20

    print("=" * 80)
    print(f" {NUM_RUNS} DQN")
    print("=" * 80)

    for run in range(1, NUM_RUNS + 1):
        try:
            run_single_experiment(run, results_save_dir)
        except Exception as e:
            print(f"\nRun  {run}  run: {str(e)}")
            import traceback
            traceback.print_exc()
            continue

    print("\n" + "=" * 80)
    print(f" {NUM_RUNS}  runcompleted")
    print(f": {results_save_dir}")
    print("=" * 80)