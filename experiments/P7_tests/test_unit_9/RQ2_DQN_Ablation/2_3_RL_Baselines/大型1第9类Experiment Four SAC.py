import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
from torch.distributions import Normal
import numpy as np
import random
import time
import math
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# === 目标路径组（与 section9 规则编号 1~160 匹配） ===
targetPaths = [
    {1, 2, 3, 4, 5, 6, 11, 12, 14, 16, 17, 18, 19, 21, 22, 23, 25, 26, 29, 30, 31, 34, 35, 36, 37, 38, 39, 40, 41, 42,
     47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100,
     101, 102, 103, 104, 108, 112, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137,
     138, 141, 142, 143, 144, 145, 146, 147, 148, 149, 157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 11, 12, 14, 15, 17, 18, 19, 23, 25, 26, 29, 30, 32, 34, 35, 36, 37, 38, 39, 40, 41, 42, 47, 48,
     49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102,
     103, 104, 108, 112, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 141,
     142, 143, 144, 145, 146, 147, 148, 149, 157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 11, 12, 14, 16, 17, 18, 19, 23, 25, 26, 29, 30, 31, 34, 35, 36, 37, 38, 39, 40, 41, 42, 47, 48,
     49, 50, 53, 54, 55, 56, 57, 58, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102, 103,
     104, 108, 111, 112, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 141,
     142, 143, 144, 145, 146, 147, 148, 149, 157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 14, 16, 17, 18, 19, 22, 25, 26, 29, 30, 31, 34, 35, 36, 37, 38, 39, 40, 41, 42, 47, 48, 49, 50,
     51, 52, 54, 55, 56, 57, 58, 63, 64, 67, 68, 69, 70, 73, 75, 76, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102,
     103, 104, 108, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 139, 140,
     141, 142, 143, 144, 145, 146, 149, 152, 154, 159, 160},

    {1, 2, 3, 4, 5, 6, 14, 16, 17, 18, 19, 21, 22, 23, 25, 26, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 47,
     48, 49, 50, 53, 54, 55, 56, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104,
     108, 111, 112, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 139, 140,
     141, 142, 143, 144, 145, 146, 149, 152, 154, 159, 160},

    {1, 2, 3, 4, 5, 6, 9, 10, 11, 12, 14, 17, 18, 19, 20, 21, 22, 25, 26, 27, 28, 29, 30, 32, 34, 35, 36, 37, 38, 39,
     40, 41, 42, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 85, 86, 87, 88, 95, 96,
     97, 98, 99, 100, 103, 104, 105, 108, 111, 112, 115, 116, 119, 120, 125, 126, 127, 128, 129, 130, 131, 132, 137,
     138, 141, 142, 143, 144, 145, 146, 157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 11, 12, 14, 17, 18, 19, 21, 22, 23, 25, 26, 29, 30, 31, 34, 35, 36, 37, 38, 39, 40, 41, 42, 47,
     48, 49, 50, 51, 52, 54, 55, 56, 57, 58, 59, 60, 63, 64, 67, 68, 69, 70, 73, 77, 78, 79, 80, 95, 96, 97, 98, 99,
     100, 103, 104, 108, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 141,
     142, 143, 144, 145, 146, 147, 148, 149, 159, 160},

    {1, 2, 3, 4, 5, 6, 9, 10, 14, 16, 17, 18, 19, 20, 21, 22, 25, 26, 29, 30, 31, 34, 35, 36, 37, 38, 39, 40, 47, 48,
     49, 50, 51, 52, 55, 56, 57, 58, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102, 103,
     104, 105, 108, 111, 112, 115, 116, 119, 120, 121, 122, 127, 128, 129, 130, 131, 132, 137, 138, 139, 140, 141, 142,
     143, 144, 145, 146, 152, 153, 157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 13, 16, 17, 18, 21, 22, 25, 26, 29, 30, 32, 34, 35, 36, 37, 38, 39, 40, 49, 50, 51,
     52, 55, 56, 57, 58, 63, 64, 69, 70, 77, 78, 79, 80, 87, 88, 93, 94, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104,
     105, 108, 111, 112, 115, 116, 119, 120, 121, 122, 127, 128, 129, 130, 131, 132, 137, 138, 139, 140, 141, 142, 143,
     144, 145, 146, 152, 153, 157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 16, 17, 18, 19, 20, 21, 22, 25, 26, 29, 30, 32, 33, 34, 35, 36, 37, 38, 49, 50, 51,
     52, 53, 54, 55, 56, 57, 58, 63, 64, 69, 70, 77, 78, 79, 80, 89, 90, 95, 96, 97, 98, 99, 100, 103, 104, 105, 108,
     111, 112, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 137, 138, 139, 140, 141, 142, 143,
     144, 145, 146, 152, 153, 157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 14, 16, 17, 18, 19, 21, 22, 23, 25, 26, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40, 47, 48, 49,
     50, 53, 54, 55, 56, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 84, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 108,
     111, 112, 115, 116, 119, 120, 121, 122, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 139, 140, 141, 142, 143,
     144, 145, 146, 149, 152, 154, 159, 160},

    {1, 2, 3, 4, 5, 6, 11, 12, 14, 16, 17, 18, 19, 21, 22, 23, 24, 25, 26, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40,
     41, 42, 47, 48, 49, 50, 54, 55, 56, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104,
     108, 111, 112, 115, 116, 119, 120, 121, 122, 125, 126, 129, 130, 131, 132, 135, 136, 137, 138, 141, 142, 143, 144,
     145, 146, 147, 148, 149, 159, 160},

    {1, 2, 3, 4, 5, 6, 14, 16, 17, 18, 19, 21, 22, 23, 24, 25, 26, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42,
     43, 44, 47, 48, 49, 50, 54, 55, 56, 67, 68, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102, 103,
     104, 108, 112, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 131, 132, 135, 136, 137, 138, 139, 140, 141, 142,
     145, 146, 150, 151, 159, 160},

    {1, 2, 3, 4, 9, 10, 11, 12, 13, 15, 17, 18, 25, 26, 27, 28, 29, 30, 32, 34, 35, 36, 37, 38, 39, 40, 49, 50, 51, 52,
     55, 56, 57, 58, 63, 64, 69, 70, 77, 78, 79, 80, 81, 82, 85, 86, 87, 88, 95, 96, 97, 98, 99, 100, 101, 102, 103,
     104, 105, 108, 111, 112, 115, 116, 119, 120, 121, 122, 127, 128, 129, 130, 131, 132, 137, 138, 141, 142, 143, 144,
     145, 146, 157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 14, 16, 17, 18, 19, 22, 23, 24, 25, 26, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40, 47, 48, 49,
     50, 53, 54, 55, 56, 69, 70, 73, 74, 77, 78, 79, 80, 83, 84, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 108, 111,
     112, 115, 116, 119, 120, 121, 122, 129, 130, 131, 132, 135, 136, 137, 138, 139, 140, 141, 142, 143, 144, 145, 146,
     149, 152, 154, 159, 160},

    {1, 2, 3, 4, 5, 6, 9, 10, 11, 12, 13, 15, 17, 18, 23, 25, 26, 27, 28, 32, 33, 34, 35, 36, 37, 38, 49, 50, 55, 56,
     57, 58, 63, 64, 69, 70, 77, 78, 79, 80, 81, 82, 85, 86, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 105, 108, 111,
     112, 113, 114, 115, 116, 119, 120, 121, 122, 127, 128, 129, 130, 131, 132, 137, 138, 141, 142, 143, 144, 145, 146,
     157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 13, 16, 17, 18, 22, 23, 24, 25, 26, 32, 33, 34, 35, 36, 49, 50, 55, 56, 63, 64, 69,
     70, 77, 78, 79, 80, 87, 88, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 108, 111, 112, 113, 114,
     115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 137, 138, 139, 140, 141, 142, 143, 144, 145, 146, 152,
     153, 159, 160},

    {1, 2, 3, 4, 5, 6, 14, 16, 19, 22, 23, 24, 25, 26, 29, 30, 31, 33, 34, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47,
     48, 49, 50, 54, 55, 56, 67, 68, 69, 70, 77, 78, 79, 80, 97, 98, 99, 100, 103, 104, 108, 111, 112, 119, 120, 121,
     122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 139, 140, 141, 142, 145, 146, 150, 151, 159, 160},

    {1, 2, 5, 6, 14, 16, 17, 18, 19, 22, 23, 24, 25, 26, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40, 47, 48, 49, 50, 53,
     54, 69, 70, 73, 74, 77, 78, 79, 80, 83, 84, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 108, 111, 112, 115, 116,
     119, 120, 121, 122, 131, 132, 135, 136, 137, 138, 139, 140, 141, 142, 145, 146, 150, 151, 155, 156, 159, 160},

    {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 13, 16, 17, 18, 22, 23, 24, 25, 26, 32, 33, 34, 35, 36, 53, 54, 55, 56, 63, 64, 69,
     70, 77, 78, 79, 80, 87, 88, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 107, 113, 114, 115, 116,
     118, 119, 120, 121, 122, 125, 126, 127, 128, 137, 138, 139, 140, 141, 142, 145, 146, 151, 155, 156, 159, 160},

    {1, 2, 5, 6, 9, 10, 14, 17, 18, 19, 20, 22, 23, 27, 28, 29, 30, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 49, 50, 54,
     55, 56, 57, 58, 63, 64, 69, 70, 73, 74, 79, 80, 87, 88, 95, 96, 97, 98, 99, 100, 103, 104, 108, 109, 111, 112, 115,
     116, 119, 120, 125, 126, 127, 128, 129, 130, 131, 132, 137, 138, 143, 144, 145, 146, 152, 153, 157, 158},

    {1, 2, 3, 4, 5, 6, 14, 16, 19, 22, 23, 24, 25, 26, 29, 30, 31, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49,
     50, 54, 55, 56, 57, 58, 61, 62, 65, 67, 68, 69, 70, 97, 98, 99, 100, 103, 104, 108, 109, 111, 112, 119, 120, 125,
     126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 139, 140, 141, 142, 145, 146, 150, 151, 159, 160},

    {1, 2, 3, 4, 5, 6, 14, 16, 19, 21, 22, 23, 24, 25, 26, 29, 30, 31, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48,
     49, 50, 54, 55, 56, 65, 69, 70, 97, 98, 99, 100, 103, 104, 108, 109, 111, 112, 119, 120, 121, 122, 125, 126, 127,
     128, 129, 130, 131, 132, 134, 135, 136, 137, 138, 141, 142, 145, 146, 147, 148, 150, 159, 160},

    {1, 2, 3, 4, 5, 6, 7, 8, 13, 16, 17, 18, 22, 23, 24, 25, 26, 32, 35, 36, 53, 54, 69, 70, 79, 80, 87, 88, 91, 92, 93,
     94, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 109, 111, 112, 115, 116, 117, 118, 119, 120, 121, 122, 125, 126,
     137, 138, 139, 140, 141, 142, 145, 146, 151, 155, 156, 159, 160},

    {1, 2, 3, 4, 5, 6, 9, 10, 11, 12, 14, 15, 17, 18, 19, 20, 23, 24, 25, 26, 32, 33, 34, 35, 36, 39, 40, 49, 50, 53,
     54, 69, 70, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 108, 109, 111, 112, 115, 116, 119, 120, 121, 122,
     123, 124, 129, 130, 137, 138, 141, 142, 145, 146, 159, 160},

    {3, 4, 5, 6, 14, 16, 19, 21, 22, 31, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 54, 56, 57, 58, 61, 62, 65, 66,
     67, 68, 69, 70, 79, 80, 99, 100, 103, 104, 108, 109, 110, 111, 112, 119, 120, 125, 126, 127, 128, 129, 130, 131,
     132, 134, 135, 136, 137, 138, 141, 142, 147, 148},

    {1, 2, 5, 6, 17, 18, 19, 22, 27, 28, 33, 34, 35, 36, 51, 52, 54, 56, 57, 58, 63, 64, 65, 69, 70, 71, 72, 89, 90, 95,
     96, 97, 98, 99, 100, 103, 104, 105, 106, 109, 112, 115, 116, 119, 120, 125, 126, 127, 128, 129, 130, 134, 143, 144,
     145, 146, 149, 152, 157, 158},

    {1, 2, 3, 4, 14, 19, 21, 22, 31, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 54, 57, 58, 59, 60, 61, 62, 65, 66, 75, 76,
     77, 78, 79, 80, 99, 100, 103, 104, 110, 111, 112, 119, 120, 125, 126, 127, 128, 129, 130, 133, 134, 135, 136, 137,
     138, 141, 142, 147, 148},
]

# === 配置（新范围：X速度 140~240，Y扭矩 200~250，Z电流 1~90） ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'ACTION_DIM': 3,
    'MIN_VALUES': np.array([140.0, 200.0, 1.0], dtype=np.float32),
    'MAX_VALUES': np.array([240.0, 250.0, 90.0], dtype=np.float32),
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 3,
    'REPLAY_BATCH_SIZE': 64,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'ACTOR_LR': 3e-4,
    'CRITIC_LR': 3e-4,
    'ALPHA_LR': 3e-4,
    'GAMMA': 0.99,
    'TAU': 0.005,
    'NUM_RUNS': 20,
    'TOP_K_SAMPLES': 20,
    'TARGET_PATHS': targetPaths
}

# === section9 触发函数（规则编号 1~160，与新范围匹配） ===
def section9_hybrid_speed_torque_current(x, y, z):
    triggered = set()
    speed_std = (1000, 2000)
    torque_std = (100, 200)
    current_std = (10, 30)

    if [(160 < x < 240)] != [(160 < x * 8 < 240)]:
        triggered.add(1)
    if [(160 < x < 240)] != [(160 < 100 < 240)]:
        triggered.add(2)

    if [(210 < y < 250)] != [(210 < y * 9 < 250)]:
        triggered.add(3)
    if [(210 < y < 250)] != [(210 < y * 12 < 250)]:
        triggered.add(4)

    if [(40 < z < 90)] != [(40 < z * 8 < 90)]:
        triggered.add(5)
    if [(40 < z < 90)] != [(40 < z * 9 < 90)]:
        triggered.add(6)

    if [(x > 190 and x < 210 and y > 225 and y < 235)] != [(x > 190 and x < 210 and y > 225 and 200 < 235)]:
        triggered.add(7)
    if [(x > 190 and x < 210 and y > 225 and y < 235)] != [(x > 190 and x < 210 and y > 225 and 220 < 235)]:
        triggered.add(8)

    if [(x > 185 and x < 215 and z > 60 and z < 70)] != [(x > 185 and x < 215 and z > 60 and 60 < 70)]:
        triggered.add(9)
    if [(x > 185 and x < 215 and z > 60 and z < 70)] != [(x > 185 and x < 215 and z > 60 and 55 < 70)]:
        triggered.add(10)

    if [(y > 223 and y < 237 and z > 60 and z < 70)] != [(y > 223 and y < 237 and z > 60 and 60 < 70)]:
        triggered.add(11)
    if [(y > 223 and y < 237 and z > 60 and z < 70)] != [(y > 223 and y < 237 and z > 60 and 67 < 70)]:
        triggered.add(12)

    if [(x > 198 and x < 202)] != [(x > 198 and x * 8 < 202)]:
        triggered.add(13)
    if [(x > 198 and x < 202)] != [(x > 198 and 200 < 202)]:
        triggered.add(14)

    if [(y > 228 and y < 232)] != [(y > 228 and y * 8 < 232)]:
        triggered.add(15)
    if [(y > 228 and y < 232)] != [(y > 228 and 200 < 232)]:
        triggered.add(16)

    if [(z > 63 and z < 67)] != [(z > 63 and 44 < 67)]:
        triggered.add(17)
    if [(z > 63 and z < 67)] != [(z > 63 and 56 < 67)]:
        triggered.add(18)

    if [(abs(x - 200) < 5)] != [(abs(x - 200) < 51)]:
        triggered.add(19)
    if [(abs(x - 200) < 5)] != [(abs(x - 200) < 15)]:
        triggered.add(20)

    if [(abs(y - 230) < 3)] != [(abs(y - 230) < 13)]:
        triggered.add(21)
    if [(abs(y - 230) < 3)] != [(abs(y - 230) < 31)]:
        triggered.add(22)

    if [(abs(z - 65) < 2)] != [(abs(z - 65) < 21)]:
        triggered.add(23)
    if [(abs(z - 65) < 2)] != [(abs(z - 65) < 12)]:
        triggered.add(24)

    if [(x > 185 and y > 223 and z > 55)] != [(x > 185 and y > 223 and z > 155)]:
        triggered.add(25)
    if [(x > 185 and y > 223 and z > 55)] != [(x > 185 and y > 223 and z > 515)]:
        triggered.add(26)

    if [(x < 215 and y < 237 and z < 75)] != [(x < 215 and y < 237 and z < 715)]:
        triggered.add(27)
    if [(x < 215 and y < 237 and z < 75)] != [(x < 215 and y < 237 and z < 225)]:
        triggered.add(28)

    if [((x + y / 10 + z) > 288 and (x + y / 10 + z) < 308)] != [((x + y / 10 + z) > 288 and (x + y / 10 + z) < 508)]:
        triggered.add(29)
    if [((x + y / 10 + z) > 288 and (x + y / 10 + z) < 308)] != [((x + y / 10 + z) > 288 and (x + y / 10 + z) < 708)]:
        triggered.add(30)

    if [(x * y / 100 > 440 and x * y / 100 < 500)] != [(x * y / 100 > 440 and x * y / 100 < 900)]:
        triggered.add(31)
    if [(x * y / 100 > 440 and x * y / 100 < 500)] != [(x * y / 100 > 440 and x * y / 100 < 440)]:
        triggered.add(32)

    if [(x * z > 12500 and x * z < 14500)] != [(x * z > 12500 and x * z < 17600)]:
        triggered.add(33)
    if [(x * z > 12500 and x * z < 14500)] != [(x * z > 12500 and x * z < 54500)]:
        triggered.add(34)

    if [(y / 10 * z > 1400 and y / 10 * z < 1600)] != [(y / 10 * z > 1400 and y / 10 * z < 7800)]:
        triggered.add(35)
    if [(y / 10 * z > 1400 and y / 10 * z < 1600)] != [(y / 10 * z > 1400 and y / 10 * z < 5600)]:
        triggered.add(36)

    if [((x + y / 10 + z) / 3 > 96 and (x + y / 10 + z) / 3 < 102)] != [((x + y / 10 + z) / 3 > 96 and (x + y / 10 + z) / 3 < 1052)]:
        triggered.add(37)
    if [((x + y / 10 + z) / 3 > 96 and (x + y / 10 + z) / 3 < 102)] != [((x + y / 10 + z) / 3 > 96 and (x + y / 10 + z) / 3 < 1502)]:
        triggered.add(38)

    if [(math.sqrt(x ** 2 + (y / 10) ** 2 + z ** 2) > 220)] != [(math.sqrt(x ** 2 + (y / 10) ** 2 + z ** 2) > 2520)]:
        triggered.add(39)
    if [(math.sqrt(x ** 2 + (y / 10) ** 2 + z ** 2) > 220)] != [(math.sqrt(x ** 2 + (y / 10) ** 2 + z ** 2) > 280)]:
        triggered.add(40)

    if [(x / (y / 10) > 8.2 and x / (y / 10) < 9.2)] != [(x / (y / 10) > 8.2 and x / (y / 10) < 92)]:
        triggered.add(41)
    if [(x / (y / 10) > 8.2 and x / (y / 10) < 9.2)] != [(x / (y / 10) > 8.2 and x / (y / 10) < 19.2)]:
        triggered.add(42)

    if [(x / z > 2.8 and x / z < 3.4)] != [(x / z > 2.8 and x / z < 13.4)]:
        triggered.add(43)
    if [(x / z > 2.8 and x / z < 3.4)] != [(x / z > 2.8 and x / z < 31.4)]:
        triggered.add(44)

    if [((y / 10) / z > 0.32 and (y / 10) / z < 0.38)] != [((y / 10) / z > 0.32 and (y / 10) / z < 38)]:
        triggered.add(45)
    if [((y / 10) / z > 0.32 and (y / 10) / z < 0.38)] != [((y / 10) / z > 0.32 and (y / 10) / z < 3.8)]:
        triggered.add(46)

    if [((x - 180) + (y - 220) / 10 > 22 and (x - 180) + (y - 220) / 10 < 32)] != [((x - 180) + (y - 220) / 10 > 22 and (x - 180) + (y - 220) / 10 < 312)]:
        triggered.add(47)
    if [((x - 180) + (y - 220) / 10 > 22 and (x - 180) + (y - 220) / 10 < 32)] != [((x - 180) + (y - 220) / 10 > 22 and (x - 180) + (y - 220) / 10 < 132)]:
        triggered.add(48)

    if [((x - 180) + (z - 50) > 35 and (x - 180) + (z - 50) < 45)] != [((x - 180) + (z - 50) > 35 and (x - 180) + (z - 50) < 415)]:
        triggered.add(49)
    if [((x - 180) + (z - 50) > 35 and (x - 180) + (z - 50) < 45)] != [((x - 180) + (z - 50) > 35 and (x - 180) + (z - 50) < 145)]:
        triggered.add(50)

    if [((y - 220) / 10 + (z - 50) > 25 and (y - 220) / 10 + (z - 50) < 35)] != [((y - 220) / 10 + (z - 50) > 25 and (y - 220) / 10 + (z - 50) < 315)]:
        triggered.add(51)
    if [((y - 220) / 10 + (z - 50) > 25 and (y - 220) / 10 + (z - 50) < 35)] != [((y - 220) / 10 + (z - 50) > 25 and (y - 220) / 10 + (z - 50) < 135)]:
        triggered.add(52)

    if [(abs((x - 200) - (y - 230) / 2) < 8)] != [(abs((x - 200) - (y - 230) / 2) < 18)]:
        triggered.add(53)
    if [(abs((x - 200) - (y - 230) / 2) < 8)] != [(abs((x - 200) - (y - 230) / 2) < 81)]:
        triggered.add(54)

    if [(abs((x - 200) - (z - 65) * 3) < 12)] != [(abs((x - 200) - (z - 65) * 3) < 92)]:
        triggered.add(55)
    if [(abs((x - 200) - (z - 65) * 3) < 12)] != [(abs((x - 200) - (z - 65) * 3) < 122)]:
        triggered.add(56)

    if [(abs((y - 230) / 10 - (z - 65) / 5) < 3)] != [(abs((y - 230) / 10 - (z - 65) / 5) < 13)]:
        triggered.add(57)
    if [(abs((y - 230) / 10 - (z - 65) / 5) < 3)] != [(abs((y - 230) / 10 - (z - 65) / 5) < 31)]:
        triggered.add(58)

    if [(x / (y / 10 + 5) > 6.5 and x / (y / 10 + 5) < 8.5)] != [(x / (y / 10 + 5) > 6.5 and x / (y / 10 + 5) < 85)]:
        triggered.add(59)
    if [(x / (y / 10 + 5) > 6.5 and x / (y / 10 + 5) < 8.5)] != [(x / (y / 10 + 5) > 6.5 and x / (y / 10 + 5) < 81.5)]:
        triggered.add(60)

    if [((y / 10) / (z + 15) > 0.26 and (y / 10) / (z + 15) < 0.34)] != [((y / 10) / (z + 15) > 0.26 and (y / 10) / (z + 15) < 10.34)]:
        triggered.add(61)
    if [((y / 10) / (z + 15) > 0.26 and (y / 10) / (z + 15) < 0.34)] != [((y / 10) / (z + 15) > 0.26 and (y / 10) / (z + 15) < 40.34)]:
        triggered.add(62)

    if [(z / (x / 5) > 1.4 and z / (x / 5) < 1.8)] != [(z / (x / 5) > 1.4 and z / (x / 5) < 11.8)]:
        triggered.add(63)
    if [(z / (x / 5) > 1.4 and z / (x / 5) < 1.8)] != [(z / (x / 5) > 1.4 and z / (x / 5) < 41.8)]:
        triggered.add(64)

    if [((x * y * z / 10000) > 28 and (x * y * z / 10000) < 36)] != [((x * y * z / 10000) > 28 and (x * y * z / 10000) < 332)]:
        triggered.add(65)
    if [((x * y * z / 10000) > 28 and (x * y * z / 10000) < 36)] != [((x * y * z / 10000) > 28 and (x * y * z / 10000) < 236)]:
        triggered.add(66)

    if [(x * 0.5 + y / 10 * 0.3 + z * 0.2 > 125 and x * 0.5 + y / 10 * 0.3 + z * 0.2 < 135)] != [(x * 0.5 + y / 10 * 0.3 + z * 0.2 > 125 and x * 0.5 + y / 10 * 0.3 + z * 0.2 < 935)]:
        triggered.add(67)
    if [(x * 0.5 + y / 10 * 0.3 + z * 0.2 > 125 and x * 0.5 + y / 10 * 0.3 + z * 0.2 < 135)] != [(x * 0.5 + y / 10 * 0.3 + z * 0.2 > 125 and x * 0.5 + y / 10 * 0.3 + z * 0.2 < 635)]:
        triggered.add(68)

    if [((x / 200) ** 0.4 * (y / 230) ** 0.3 * (z / 65) ** 0.3 > 0.92)] != [((x / 200) ** 0.4 * (y / 230) ** 0.3 * (z / 65) ** 0.3 > 9.2)]:
        triggered.add(69)
    if [((x / 200) ** 0.4 * (y / 230) ** 0.3 * (z / 65) ** 0.3 > 0.92)] != [((x / 200) ** 0.4 * (y / 230) ** 0.3 * (z / 65) ** 0.3 > 92)]:
        triggered.add(70)

    if [((x - 200) * (y - 230) / 10 > -80 and (x - 200) * (y - 230) / 10 < 80)] != [((x - 200) * (y - 230) / 10 > -80 and (x - 200) * (y - 230) / 10 < 180)]:
        triggered.add(71)
    if [((x - 200) * (y - 230) / 10 > -80 and (x - 200) * (y - 230) / 10 < 80)] != [((x - 200) * (y - 230) / 10 > -80 and (x - 200) * (y - 230) / 10 < 810)]:
        triggered.add(72)

    if [((x - 200) * (z - 65) > -120 and (x - 200) * (z - 65) < 120)] != [((x - 200) * (z - 65) > -120 and (x - 200) * (z - 65) < 820)]:
        triggered.add(73)
    if [((x - 200) * (z - 65) > -120 and (x - 200) * (z - 65) < 120)] != [((x - 200) * (z - 65) > -120 and (x - 200) * (z - 65) < 520)]:
        triggered.add(74)

    if [((y - 230) / 10 * (z - 65) > -30 and (y - 230) / 10 * (z - 65) < 30)] != [((y - 230) / 10 * (z - 65) > -30 and (y - 230) / 10 * (z - 65) < 310)]:
        triggered.add(75)
    if [((y - 230) / 10 * (z - 65) > -30 and (y - 230) / 10 * (z - 65) < 30)] != [((y - 230) / 10 * (z - 65) > -30 and (y - 230) / 10 * (z - 65) < 130)]:
        triggered.add(76)

    if [(abs(x / 200 + y / 230 + z / 65 - 3) < 0.18)] != [(abs(x / 200 + y / 230 + z / 65 - 3) < 18)]:
        triggered.add(77)
    if [(abs(x / 200 + y / 230 + z / 65 - 3) < 0.18)] != [(abs(x / 200 + y / 230 + z / 65 - 3) < 10.18)]:
        triggered.add(78)

    if [(abs((x / 200) * (y / 230) * (z / 65) - 1) < 0.12)] != [(abs((x / 200) * (y / 230) * (z / 65) - 1) < 10.12)]:
        triggered.add(79)
    if [(abs((x / 200) * (y / 230) * (z / 65) - 1) < 0.12)] != [(abs((x / 200) * (y / 230) * (z / 65) - 1) < 12.12)]:
        triggered.add(80)

    if [(x > 197 and x < 203 and y > 228 and y < 232 and z > 63.5 and z < 66.5)] != [(x > 197 and x < 203 and y > 228 and y < 232 and z > 63.5 and z < 616.5)]:
        triggered.add(81)
    if [(x > 197 and x < 203 and y > 228 and y < 232 and z > 63.5 and z < 66.5)] != [(x > 197 and x < 203 and y > 228 and y < 232 and z > 63.5 and z < 166.5)]:
        triggered.add(82)

    if [(abs(x / (y / 10) - 8.7) < 0.3 and abs(z - 65) < 1.5)] != [(abs(x / (y / 10) - 8.7) < 0.3 and abs(z - 65) < 11.5)]:
        triggered.add(83)
    if [(abs(x / (y / 10) - 8.7) < 0.3 and abs(z - 65) < 1.5)] != [(abs(x / (y / 10) - 8.7) < 0.3 and abs(z - 65) < 15)]:
        triggered.add(84)

    if [(x * y / 100 > 455 and x * y / 100 < 465 and z > 64 and z < 66)] != [(x * y / 100 > 455 and x * y / 100 < 465 and z > 64 and z < 166)]:
        triggered.add(85)
    if [(x * y / 100 > 455 and x * y / 100 < 465 and z > 64 and z < 66)] != [(x * y / 100 > 455 and x * y / 100 < 465 and z > 64 and z < 616)]:
        triggered.add(86)

    if [((x + y / 10) / 2 > 112 and (x + y / 10) / 2 < 116 and z > 63.5 and z < 66.5)] != [((x + y / 10) / 2 > 112 and (x + y / 10) / 2 < 116 and z > 63.5 and z < 166.5)]:
        triggered.add(87)
    if [((x + y / 10) / 2 > 112 and (x + y / 10) / 2 < 116 and z > 63.5 and z < 66.5)] != [((x + y / 10) / 2 > 112 and (x + y / 10) / 2 < 116 and z > 63.5 and z < 661.5)]:
        triggered.add(88)

    if [(abs(x - y / 10) < 175 and z > 63.5 and z < 66.5)] != [(abs(x - y / 10) < 175 and z > 63.5 and z < 616.5)]:
        triggered.add(89)
    if [(abs(x - y / 10) < 175 and z > 63.5 and z < 66.5)] != [(abs(x - y / 10) < 175 and z > 63.5 and z < 6615)]:
        triggered.add(90)

    if [(math.sqrt((x - 200) ** 2 + (y / 10 - 23) ** 2) < 3 and abs(z - 65) < 1.5)] != [(math.sqrt((x - 200) ** 2 + (y / 10 - 23) ** 2) < 3 and abs(z - 65) < 15)]:
        triggered.add(91)
    if [(math.sqrt((x - 200) ** 2 + (y / 10 - 23) ** 2) < 3 and abs(z - 65) < 1.5)] != [(math.sqrt((x - 200) ** 2 + (y / 10 - 23) ** 2) < 3 and abs(z - 65) < 11.5)]:
        triggered.add(92)

    if [(x / 200 > 0.985 and x / 200 < 1.015 and y / 230 > 0.987 and y / 230 < 1.013)] != [(x / 200 > 0.985 and x / 200 < 1.015 and y / 230 > 0.987 and y / 230 < 10.13)]:
        triggered.add(93)
    if [(x / 200 > 0.985 and x / 200 < 1.015 and y / 230 > 0.987 and y / 230 < 1.013)] != [(x / 200 > 0.985 and x / 200 < 1.015 and y / 230 > 0.987 and y / 230 < 1013)]:
        triggered.add(94)

    if [(z / 65 > 0.985 and z / 65 < 1.015)] != [(z / 65 > 0.985 and z / 65 < 121.015)]:
        triggered.add(95)
    if [(z / 65 > 0.985 and z / 65 < 1.015)] != [(z / 65 > 0.985 and z / 65 < 111.015)]:
        triggered.add(96)

    if [((x / 200 + y / 230 + z / 65) / 3 > 0.985 and (x / 200 + y / 230 + z / 65) / 3 < 1.015)] != [((x / 200 + y / 230 + z / 65) / 3 > 0.985 and (x / 200 + y / 230 + z / 65) / 3 < 2.115)]:
        triggered.add(97)
    if [((x / 200 + y / 230 + z / 65) / 3 > 0.985 and (x / 200 + y / 230 + z / 65) / 3 < 1.015)] != [((x / 200 + y / 230 + z / 65) / 3 > 0.985 and (x / 200 + y / 230 + z / 65) / 3 < 15)]:
        triggered.add(98)

    if [(max(abs(x / 200 - 1), abs(y / 230 - 1), abs(z / 65 - 1)) < 0.02)] != [(max(abs(x / 200 - 1), abs(y / 230 - 1), abs(z / 65 - 1)) < 22)]:
        triggered.add(99)
    if [(max(abs(x / 200 - 1), abs(y / 230 - 1), abs(z / 65 - 1)) < 0.02)] != [(max(abs(x / 200 - 1), abs(y / 230 - 1), abs(z / 65 - 1)) < 10.02)]:
        triggered.add(100)

    if [(min(x / 200, y / 230, z / 65) > 0.985)] != [(min(x / 200, y / 230, z / 65) > 9.85)]:
        triggered.add(101)
    if [(min(x / 200, y / 230, z / 65) > 0.985)] != [(min(x / 200, y / 230, z / 65) > 10.985)]:
        triggered.add(102)

    if [(max(x / 200, y / 230, z / 65) < 1.015)] != [(max(x / 200, y / 230, z / 65) < 10.15)]:
        triggered.add(103)
    if [(max(x / 200, y / 230, z / 65) < 1.015)] != [(max(x / 200, y / 230, z / 65) < 15)]:
        triggered.add(104)

    if [(abs(max(x, y / 10 * 8.7, z) - min(x, y / 10 * 8.7, z)) < 15)] != [(abs(max(x, y / 10 * 8.7, z) - min(x, y / 10 * 8.7, z)) < 125)]:
        triggered.add(105)
    if [(abs(max(x, y / 10 * 8.7, z) - min(x, y / 10 * 8.7, z)) < 15)] != [(abs(max(x, y / 10 * 8.7, z) - min(x, y / 10 * 8.7, z)) < 88)]:
        triggered.add(106)

    if [((x + y / 10 + z) > 297 and (x + y / 10 + z) < 299)] != [((x + y / 10 + z) > 297 and (x + y / 10 + z) < 99)]:
        triggered.add(107)
    if [((x + y / 10 + z) > 297 and (x + y / 10 + z) < 299)] != [((x + y / 10 + z) > 297 and (x + y / 10 + z) < 549)]:
        triggered.add(108)

    if [((x * y * z / 10000) > 31.5 and (x * y * z / 10000) < 345)] != [((x * y * z / 10000) > 31.5 and (x * y * z / 10000) < 132.5)]:
        triggered.add(109)
    if [((x * y * z / 10000) > 31.5 and (x * y * z / 10000) < 32.5)] != [((x * y * z / 10000) > 31.5 and (x * y * z / 10000) < 232.5)]:
        triggered.add(110)

    if [(abs((x + y / 10 + z) / 3 - 99.3) < 0.8)] != [(abs((x + y / 10 + z) / 3 - 99.3) < 8)]:
        triggered.add(111)
    if [(abs((x + y / 10 + z) / 3 - 99.3) < 0.8)] != [(abs((x + y / 10 + z) / 3 - 99.3) < 10.8)]:
        triggered.add(112)

    if [(x > 199 and x < 201 and y > 229 and y < 231)] != [(x > 199 and x < 201 and y > 229 and y < 2231)]:
        triggered.add(113)
    if [(x > 199 and x < 201 and y > 229 and y < 231)] != [(x > 199 and x < 201 and y > 229 and y < 1231)]:
        triggered.add(114)

    if [(z > 64.5 and z < 65.5)] != [(z > 64.5 and z < 615.5)]:
        triggered.add(115)
    if [(z > 64.5 and z < 65.5)] != [(z > 64.5 and z < 655)]:
        triggered.add(116)

    if [((x - 200) ** 2 + (y / 10 - 23) ** 2 + (z - 65) ** 2) < 9] != [((x - 200) ** 2 + (y / 10 - 23) ** 2 + (z - 65) ** 2) < 19]:
        triggered.add(117)
    if [((x - 200) ** 2 + (y / 10 - 23) ** 2 + (z - 65) ** 2) < 9] != [((x - 200) ** 2 + (y / 10 - 23) ** 2 + (z - 65) ** 2) < 91]:
        triggered.add(118)

    if [(abs(x / (y / 10) / z - 0.134) < 0.005)] != [(abs(x / (y / 10) / z - 0.134) < 10.005)]:
        triggered.add(119)
    if [(abs(x / (y / 10) / z - 0.134) < 0.005)] != [(abs(x / (y / 10) / z - 0.134) < 5)]:
        triggered.add(120)

    if [(x < 188 or y < 225 or z < 58)] != [(x < 188 or y < 225 or z < 158)]:
        triggered.add(121)
    if [(x < 188 or y < 225 or z < 58)] != [(x < 188 or y < 225 or 28 < 58)]:
        triggered.add(122)

    if [(x > 212 or y > 235 or z > 72)] != [(x > 212 or y > 235 or 12 * z > 72)]:
        triggered.add(123)
    if [(x > 212 or y > 235 or z > 72)] != [(x > 212 or y > 235 or 88 > 72)]:
        triggered.add(124)

    if [(abs(x / (y / 10) - 8.7) > 0.5)] != [(abs(x / (y / 10) - 8.7) > 10.5)]:
        triggered.add(125)
    if [(abs(x / (y / 10) - 8.7) > 0.5)] != [(abs(x / (y / 10) - 8.7) > 5)]:
        triggered.add(126)

    if [(abs(x / z - 3.08) > 0.3)] != [(abs(x / z - 3.08) > 10.3)]:
        triggered.add(127)
    if [(abs(x / z - 3.08) > 0.3)] != [(abs(x / z - 3.08) > 123)]:
        triggered.add(128)

    if [(abs((y / 10) / z - 0.354) > 0.03)] != [(abs((y / 10) / z - 0.354) > 3)]:
        triggered.add(129)
    if [(abs((y / 10) / z - 0.354) > 0.03)] != [(abs((y / 10) / z - 0.354) > 10.03)]:
        triggered.add(130)

    if [((x + y / 10 + z) < 292 or (x + y / 10 + z) > 304)] != [((x + y / 10 + z) < 292 or (x + y / 10 + z) > 3224)]:
        triggered.add(131)
    if [((x + y / 10 + z) < 292 or (x + y / 10 + z) > 304)] != [((x + y / 10 + z) < 292 or (x + y / 10 + z) > 1304)]:
        triggered.add(132)

    if [((x * y * z / 10000) < 29 or (x * y * z / 10000) > 35)] != [((x * y * z / 10000) < 29 or (x * y * z / 10000) > 135)]:
        triggered.add(133)
    if [((x * y * z / 10000) < 29 or (x * y * z / 10000) > 35)] != [((x * y * z / 10000) < 29 or (x * y * z / 10000) > 315)]:
        triggered.add(134)

    if [(x > 215)] != [(x > 1215)]:
        triggered.add(135)
    if [(x > 215)] != [(x > 2115)]:
        triggered.add(136)

    if [(x < 185)] != [(x < 1815)]:
        triggered.add(137)
    if [(x < 185)] != [(x < 1185)]:
        triggered.add(138)

    if [(y > 237)] != [(y > 2137)]:
        triggered.add(139)
    if [(y > 237)] != [(y > 2317)]:
        triggered.add(140)

    if [(y < 223)] != [(y < 2123)]:
        triggered.add(141)
    if [(y < 223)] != [(y < 2283)]:
        triggered.add(142)

    if [(z > 75)] != [(z > 751)]:
        triggered.add(143)
    if [(z > 75)] != [(z > 175)]:
        triggered.add(144)

    if [(z < 55)] != [(z < 515)]:
        triggered.add(145)
    if [(z < 55)] != [(z < 559)]:
        triggered.add(146)

    if [(x > 215 or x < 185) and (y > 237 or y < 223)] != [(x > 215 or x < 185) and (y > 237 or y < 2213)]:
        triggered.add(147)
    if [(x > 215 or x < 185) and (y > 237 or y < 223)] != [(x > 215 or x < 185) and (y > 237 or 200 < 223)]:
        triggered.add(148)

    if [(x > 215 or x < 185) and (z > 75 or z < 55)] != [(x > 215 or x < 185) and (z > 175 or z < 55)]:
        triggered.add(149)
    if [(x > 215 or x < 185) and (z > 75 or z < 55)] != [(x > 215 or x < 185) and (z > 75 or z < 551)]:
        triggered.add(150)

    if [(y > 237 or y < 223) and (z > 75 or z < 55)] != [(y > 237 or y < 223) and (z > 75 or z < 515)]:
        triggered.add(151)
    if [(y > 237 or y < 223) and (z > 75 or z < 55)] != [(y > 237 or y < 223) and (z > 175 or z < 55)]:
        triggered.add(152)

    if [(x > 215 or x < 185) and (y > 237 or y < 223) and (z > 75 or z < 55)] != [(x > 215 or x < 1185) and (y > 237 or y < 223) and (z > 75 or z < 55)]:
        triggered.add(153)
    if [(x > 215 or x < 185) and (y > 237 or y < 223) and (z > 75 or z < 55)] != [(x > 215 or x < 185) and (y > 2317 or y < 223) and (z > 75 or z < 55)]:
        triggered.add(154)

    if [(abs(x - 200) > 20) or (abs(y - 230) > 15) or (abs(z - 65) > 10)] != [(abs(x - 200) > 20) or (abs(y - 230) > 115) or (abs(z - 65) > 10)]:
        triggered.add(155)
    if [(abs(x - 200) > 20) or (abs(y - 230) > 15) or (abs(z - 65) > 10)] != [(abs(x - 200) > 20) or (abs(y - 230) > 415) or (abs(z - 65) > 10)]:
        triggered.add(156)

    if [(x > 220) or (y > 240) or (z > 80)] != [(x > 220) or (y > 240) or (z > 810)]:
        triggered.add(157)
    if [(x > 220) or (y > 240) or (z > 80)] != [(x > 220) or (y > 240) or (z > 911)]:
        triggered.add(158)
    if [(x < 175) or (y < 215) or (z < 45)] != [(x < 175) or (y < 215) or (z < 415)]:
        triggered.add(159)
    if [(x < 175) or (y < 215) or (z < 45)] != [(x < 175) or (y < 215) or (z < 145)]:
        triggered.add(160)
    return triggered

# === 绑定执行函数（使用 section9） ===
execute_Tr = section9_hybrid_speed_torque_current

# === 状态处理辅助函数（使用新范围） ===
def clip_state(state):
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return np.clip(state, min_vals, max_vals)

def normalize_state(state):
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return 2 * (state - min_vals) / (max_vals - min_vals) - 1

def denormalize_state(normalized_state):
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return (normalized_state + 1) * (max_vals - min_vals) / 2 + min_vals

def coverage_similarity(triggered, target_path):
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0
    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)

def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']
    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']
    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']
    return reward

# === SAC Actor ===
class GaussianPolicy(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dim=256):
        super(GaussianPolicy, self).__init__()
        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.mean = nn.Linear(hidden_dim, action_dim)
        self.log_std = nn.Linear(hidden_dim, action_dim)
        # 分维度动作步长：速度±8、扭矩±3、电流±4，适配新范围量级
        self.register_buffer('action_scale', torch.tensor([8.0, 3.0, 4.0]))
        self.register_buffer('action_bias', torch.tensor([0.0, 0.0, 0.0]))

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))
        mean = self.mean(x)
        log_std = self.log_std(x)
        log_std = torch.clamp(log_std, min=-20, max=2)
        return mean, log_std

    def sample(self, state):
        mean, log_std = self.forward(state)
        std = log_std.exp()
        normal = Normal(mean, std)
        x_t = normal.rsample()
        action = torch.tanh(x_t) * self.action_scale + self.action_bias
        log_prob = normal.log_prob(x_t)
        log_prob -= torch.log(self.action_scale * (1 - torch.tanh(x_t).pow(2)) + 1e-6)
        log_prob = log_prob.sum(1, keepdim=True)
        mean = torch.tanh(mean) * self.action_scale + self.action_bias
        return action, log_prob, mean

# === SAC Critic ===
class QNetwork(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dim=256):
        super(QNetwork, self).__init__()
        self.fc1 = nn.Linear(state_dim + action_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, 1)
        self.fc4 = nn.Linear(state_dim + action_dim, hidden_dim)
        self.fc5 = nn.Linear(hidden_dim, hidden_dim)
        self.fc6 = nn.Linear(hidden_dim, 1)

    def forward(self, state, action):
        xu = torch.cat([state, action], 1)
        x1 = F.relu(self.fc1(xu))
        x1 = F.relu(self.fc2(x1))
        q1 = self.fc3(x1)
        x2 = F.relu(self.fc4(xu))
        x2 = F.relu(self.fc5(x2))
        q2 = self.fc6(x2)
        return q1, q2

# === 经验回放缓冲区 ===
class EnhancedReplayBuffer:
    def __init__(self, capacity=50000):
        self.buffer = deque(maxlen=capacity)
        self.experience_info = deque(maxlen=capacity)

    def push(self, state, action, reward, next_state, done, path_idx, similarity):
        self.buffer.append((state, action, reward, next_state, done))
        self.experience_info.append({'path_idx': path_idx, 'similarity': similarity})

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return None
        batch = random.sample(self.buffer, batch_size)
        state, action, reward, next_state, done = map(np.stack, zip(*batch))
        return (
            torch.FloatTensor(state).to(device),
            torch.FloatTensor(action).to(device),
            torch.FloatTensor(reward).unsqueeze(1).to(device),
            torch.FloatTensor(next_state).to(device),
            torch.FloatTensor(done).unsqueeze(1).to(device)
        )

    def get_top_k_per_path(self, num_paths, k=20):
        path_samples = {i: [] for i in range(num_paths)}
        for idx, info in enumerate(self.experience_info):
            path_idx = info['path_idx']
            similarity = info['similarity']
            path_samples[path_idx].append((idx, similarity, self.buffer[idx]))

        top_k_results = {}
        for path_idx in range(num_paths):
            samples = path_samples[path_idx]
            if len(samples) == 0:
                top_k_results[path_idx] = []
                continue
            samples.sort(key=lambda x: x[1], reverse=True)
            top_k = samples[:k]
            top_k_results[path_idx] = []
            for sample in top_k:
                normalized_state = sample[2][0]
                original_state = denormalize_state(normalized_state)
                # 速度、扭矩取整，电流保留1位小数
                x, y, z = round(original_state[0]), round(original_state[1]), round(original_state[2], 1)
                triggered = execute_Tr(x, y, z)
                top_k_results[path_idx].append({
                    'state': np.array([x, y, z], dtype=np.float32),
                    'similarity': sample[1],
                    'triggered': triggered
                })
        return top_k_results

    def __len__(self):
        return len(self.buffer)

# === SAC Agent ===
class SACAgent:
    def __init__(self, state_dim=3, action_dim=3):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.policy = GaussianPolicy(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.policy_optimizer = optim.Adam(self.policy.parameters(), lr=EXPERIMENT_CONFIG['ACTOR_LR'])
        self.critic = QNetwork(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_target = QNetwork(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_target.load_state_dict(self.critic.state_dict())
        self.critic_optimizer = optim.Adam(self.critic.parameters(), lr=EXPERIMENT_CONFIG['CRITIC_LR'])
        self.target_entropy = -action_dim
        self.log_alpha = torch.zeros(1, requires_grad=True, device=device)
        self.alpha_optimizer = optim.Adam([self.log_alpha], lr=EXPERIMENT_CONFIG['ALPHA_LR'])
        self.replay_buffer = EnhancedReplayBuffer()
        self.replay_train_count = 0

    def get_action(self, state, deterministic=False):
        normalized_state = normalize_state(state)
        state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)
        with torch.no_grad():
            if deterministic:
                _, _, action = self.policy.sample(state_tensor)
            else:
                action, _, _ = self.policy.sample(state_tensor)
        action = action.cpu().numpy()[0]
        return action

    def store_experience(self, state, action, reward, next_state, done, path_idx, similarity):
        normalized_state = normalize_state(state)
        normalized_next_state = normalize_state(next_state)
        self.replay_buffer.push(
            normalized_state, action, reward,
            normalized_next_state, done, path_idx, similarity
        )

    def replay_train(self):
        batch = self.replay_buffer.sample(EXPERIMENT_CONFIG['REPLAY_BATCH_SIZE'])
        if batch is None:
            return
        state, action, reward, next_state, done = batch

        with torch.no_grad():
            next_action, next_log_prob, _ = self.policy.sample(next_state)
            q1_next, q2_next = self.critic_target(next_state, next_action)
            q_next = torch.min(q1_next, q2_next)
            target_q = reward + (1 - done) * EXPERIMENT_CONFIG['GAMMA'] * (
                    q_next - self.log_alpha.exp() * next_log_prob)

        q1, q2 = self.critic(state, action)
        critic_loss = F.mse_loss(q1, target_q) + F.mse_loss(q2, target_q)
        self.critic_optimizer.zero_grad()
        critic_loss.backward()
        self.critic_optimizer.step()

        new_action, log_prob, _ = self.policy.sample(state)
        q1_new, q2_new = self.critic(state, new_action)
        q_new = torch.min(q1_new, q2_new)
        policy_loss = (self.log_alpha.exp() * log_prob - q_new).mean()
        self.policy_optimizer.zero_grad()
        policy_loss.backward()
        self.policy_optimizer.step()

        alpha_loss = -(self.log_alpha * (log_prob + self.target_entropy).detach()).mean()
        self.alpha_optimizer.zero_grad()
        alpha_loss.backward()
        self.alpha_optimizer.step()

        for param, target_param in zip(self.critic.parameters(), self.critic_target.parameters()):
            target_param.data.copy_(
                EXPERIMENT_CONFIG['TAU'] * param.data + (1 - EXPERIMENT_CONFIG['TAU']) * target_param.data)

        self.replay_train_count += 1
        if self.replay_train_count % 2 == 0:
            alpha_value = self.log_alpha.exp().item()
            print(f"  -> SAC 更新 (第 {self.replay_train_count} 次), Alpha={alpha_value:.4f}")

# === 性能计算函数 ===
def calculate_run_performance(run_idx, sac_results, training_time, total_steps, update_count, agent):
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)
    all_similarities = []
    total_samples = 0
    total_reward = 0

    for path_idx in range(num_paths):
        samples = sac_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']
            total_reward += reward
            all_similarities.append(similarity)
            total_samples += 1

    average_reward = total_reward / total_samples if total_samples > 0 else 0
    convergence = np.mean(all_similarities) if all_similarities else 0
    environment_adaptability = 1 / (np.std(all_similarities) + 1e-8) if len(all_similarities) > 1 else 0
    generalization_ability = convergence
    computational_efficiency = total_steps / training_time if training_time > 0 else 0
    policy_update_frequency = update_count / training_time if training_time > 0 else 0
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0

    return {
        'Run': run_idx + 1,
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }

# === Excel 导出函数 ===
def export_to_excel(all_sac_results, all_performance_data, target_paths, output_path="SAC_20_run.xlsx"):
    print("\n正在导出数据到 Excel...")
    all_sac_summary_data = []
    all_sac_detailed_data = []

    for run_idx, (sac_results, performance_data) in enumerate(zip(all_sac_results, all_performance_data)):
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = sac_results[path_idx]
            if len(samples) == 0:
                all_sac_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue
            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = 'Yes' if perfect_count > 0 else 'No'
            all_sac_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })

            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']
                all_sac_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Sample ID': sample_idx + 1,
                    'X(速度)': int(state[0]),
                    'Y(扭矩)': int(state[1]),
                    'Z(电流)': float(state[2]),
                    'Similarity': round(similarity, 4),
                    'Perfect Coverage': 'Yes' if abs(similarity - 1.0) < 0.001 else 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path))),
                    'Triggered Rules': ', '.join(map(str, sorted(triggered))),
                    'Hit Rule Count': len(target_path.intersection(triggered)),
                    'Target Rule Count': len(target_path)
                })

    summary_df = pd.DataFrame(all_sac_summary_data)
    detailed_df = pd.DataFrame(all_sac_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='SACPath', index=False)
        detailed_df.to_excel(writer, sheet_name='SACDetailed Sample Data', index=False)
        selected_columns = [
            'Run', 'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df[selected_columns].to_excel(writer, sheet_name='Metric', index=False)

        workbook = writer.book
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        perfect_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        for sheet_name in ['SACPath', 'SACDetailed Sample Data', 'Metric']:
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

        ws1 = writer.sheets['SACPath']
        for row_idx in range(2, ws1.max_row + 1):
            if ws1.cell(row_idx, 9).value == 'Yes':
                for col_idx in range(1, ws1.max_column + 1):
                    ws1.cell(row_idx, col_idx).fill = perfect_fill

    print(f"文件已保存到: {output_path}")
    print(f"  - Sheet1: SACPath 统计 ({len(all_sac_summary_data)} 条)")
    print(f"  - Sheet2: SACDetailed Sample Data ({len(all_sac_detailed_data)} 条)")
    print(f"  - Sheet3: Metric ({len(all_performance_data)} 条)")

# === 单次训练流程 ===
def train_sac_workflow():
    print("=" * 80)
    print("SAC 训练 - 速度扭矩电流混合控制 (section9)")
    print("状态范围: X速度 140~240, Y扭矩 200~250, Z电流 1~90")
    print("相似度 = 命中规则数 / 目标路径规则数")
    print("=" * 80)

    agent = SACAgent()
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n生成初始样本: {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']} / 路径")
    path_samples = {}
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']

    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            state = np.random.uniform(min_vals, max_vals).astype(np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  路径 {path_idx + 1}/{num_paths}: {len(samples)} 个初始样本")

    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size

    print(f"\n批次大小: {batch_size}, 单样本迭代步数: {EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(f"总批次数: {num_batches} /路径 × {num_paths} 路径 = {num_batches * num_paths}")
    print("-" * 80)

    for batch_idx in range(num_batches):
        print(f"\n批次 {batch_idx + 1}/{num_batches}")

        for path_idx in range(num_paths):
            target_path = target_paths[path_idx]
            batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]

            batch_rewards = []
            batch_similarities = []

            for initial_state in batch_samples:
                state = initial_state.copy()
                episode_reward = 0
                final_similarity = 0

                for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                    action = agent.get_action(state, deterministic=False)
                    next_state = state + action
                    next_state = clip_state(next_state)

                    x, y, z = next_state
                    triggered = execute_Tr(x, y, z)
                    reward = unified_reward_function(triggered, target_path)
                    similarity = coverage_similarity(triggered, target_path)

                    done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)

                    agent.store_experience(
                        state, action, reward, next_state, done,
                        path_idx, similarity
                    )

                    state = next_state
                    episode_reward += reward
                    final_similarity = similarity
                    total_steps += 1

                batch_rewards.append(episode_reward)
                batch_similarities.append(final_similarity)

            avg_reward = np.mean(batch_rewards)
            avg_sim = np.mean(batch_similarities)
            print(f"  路径 {path_idx+1}: 平均奖励={avg_reward:.2f}, 平均相似度={avg_sim:.4f}")

        print("\n  执行SAC网络参数更新...")
        agent.replay_train()

    training_time = time.time() - start_time

    print("\n" + "=" * 80)
    print(f"本轮训练完成 | 总耗时: {training_time:.2f}s | 总交互步数: {total_steps}")
    print(f"网络更新次数: {agent.replay_train_count}")
    print("=" * 80)

    print(f"\n提取每条路径 Top-{EXPERIMENT_CONFIG['TOP_K_SAMPLES']} 高相似度样本...")
    top_k_results = agent.replay_buffer.get_top_k_per_path(num_paths, EXPERIMENT_CONFIG['TOP_K_SAMPLES'])

    return agent, top_k_results, training_time, total_steps, agent.replay_train_count

# === 主程序入口 ===
def main():
    print("=" * 80)
    print("SAC 速度扭矩电流混合控制 (section9) - 20轮完整实验")
    print("=" * 80)

    all_sac_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'='*80}")
        print(f"第 {run_idx+1}/{EXPERIMENT_CONFIG['NUM_RUNS']} 轮实验启动")
        print(f"{'='*80}")

        agent, sac_results, training_time, total_steps, update_count = train_sac_workflow()
        performance = calculate_run_performance(
            run_idx, sac_results, training_time, total_steps, update_count, agent
        )

        all_sac_results.append(sac_results)
        all_performance_data.append(performance)

        print(f"\n第 {run_idx+1} 轮核心指标:")
        print(f"  平均相似度: {performance['Average Similarity']:.4f}")
        print(f"  最高相似度: {performance['Max Similarity']:.4f}")
        print(f"  平均奖励: {performance['Average Reward']:.2f}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"SAC_section9_20run_{timestamp}.xlsx"
    export_to_excel(all_sac_results, all_performance_data, target_paths, output_file)

    avg_sim_list = [p['Average Similarity'] for p in all_performance_data]
    max_sim_list = [p['Max Similarity'] for p in all_performance_data]
    print("\n" + "=" * 80)
    print("20轮实验整体汇总:")
    print(f"  平均相似度 均值: {np.mean(avg_sim_list):.4f}")
    print(f"  平均相似度 标准差: {np.std(avg_sim_list):.4f}")
    print(f"  单轮最高平均相似度: {np.max(avg_sim_list):.4f}")
    print(f"  全局最高相似度: {np.max(max_sim_list):.4f}")
    print("=" * 80)


if __name__ == "__main__":
    main()