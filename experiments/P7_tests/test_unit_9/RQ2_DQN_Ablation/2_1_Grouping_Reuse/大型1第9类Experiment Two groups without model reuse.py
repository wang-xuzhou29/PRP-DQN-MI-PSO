import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import math

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === 全局取值范围配置（新范围：x:140~240, y:200~250, z:1~90） ===
MIN_X = 140
MAX_X = 240   # 温度
MIN_Y = 200
MAX_Y = 250   # 电压
MIN_Z = 1
MAX_Z = 90    # 流量

# ========== 规则触发函数（section9_hybrid_speed_torque_current） ==========
def section9_hybrid_speed_torque_current(x, y, z):
    triggered = set()
    speed_std = (1000, 2000)
    torque_std = (100, 200)
    current_std = (10, 30)

    if [(160 < x < 240)] != [
        (160 < x * 8 < 240)]:
        triggered.add(1)
    if [(160 < x < 240)] != [
        (160 < 100 < 240)]:
        triggered.add(2)

    if [(210 < y < 250)] != [
        (210 < y * 9 < 250)]:
        triggered.add(3)
    if [(210 < y < 250)] != [
        (210 < y * 12 < 250)]:
        triggered.add(4)

    if [(40 < z < 90)] != [
        (40 < z * 8 < 90)]:
        triggered.add(5)
    if [(40 < z < 90)] != [
        (40 < z * 9 < 90)]:
        triggered.add(6)

    if [(x > 190 and x < 210 and y > 225 and y < 235)] != [
        (x > 190 and x < 210 and y > 225 and 200 < 235)]:
        triggered.add(7)
    if [(x > 190 and x < 210 and y > 225 and y < 235)] != [
        (x > 190 and x < 210 and y > 225 and 220 < 235)]:
        triggered.add(8)

    if [(x > 185 and x < 215 and z > 60 and z < 70)] != [
        (x > 185 and x < 215 and z > 60 and 60 < 70)]:
        triggered.add(9)
    if [(x > 185 and x < 215 and z > 60 and z < 70)] != [
        (x > 185 and x < 215 and z > 60 and 55 < 70)]:
        triggered.add(10)

    if [(y > 223 and y < 237 and z > 60 and z < 70)] != [
        (y > 223 and y < 237 and z > 60 and 60 < 70)]:
        triggered.add(11)
    if [(y > 223 and y < 237 and z > 60 and z < 70)] != [
        (y > 223 and y < 237 and z > 60 and 67 < 70)]:
        triggered.add(12)

    if [(x > 198 and x < 202)] != [(x > 198 and x * 8 < 202)]:
        triggered.add(13)
    if [(x > 198 and x < 202)] != [(x > 198 and 200 < 202)]:
        triggered.add(14)

    if [(y > 228 and y < 232)] != [(y > 228 and y * 8 < 232)]:
        triggered.add(15)
    if [(y > 228 and y < 232)] != [(y > 228 and 200 < 232)]:
        triggered.add(16)

    if [(z > 63 and z < 67)] != [(z > 63 and 44 < 67)]:
        triggered.add(17)
    if [(z > 63 and z < 67)] != [(z > 63 and 56 < 67)]:
        triggered.add(18)

    if [(abs(x - 200) < 5)] != [(abs(x - 200) < 51)]:
        triggered.add(19)
    if [(abs(x - 200) < 5)] != [(abs(x - 200) < 15)]:
        triggered.add(20)

    if [(abs(y - 230) < 3)] != [(abs(y - 230) < 13)]:
        triggered.add(21)
    if [(abs(y - 230) < 3)] != [(abs(y - 230) < 31)]:
        triggered.add(22)

    if [(abs(z - 65) < 2)] != [(abs(z - 65) < 21)]:
        triggered.add(23)
    if [(abs(z - 65) < 2)] != [(abs(z - 65) < 12)]:
        triggered.add(24)

    if [(x > 185 and y > 223 and z > 55)] != [(x > 185 and y > 223 and z > 155)]:
        triggered.add(25)
    if [(x > 185 and y > 223 and z > 55)] != [(x > 185 and y > 223 and z > 515)]:
        triggered.add(26)

    if [(x < 215 and y < 237 and z < 75)] != [(x < 215 and y < 237 and z < 715)]:
        triggered.add(27)
    if [(x < 215 and y < 237 and z < 75)] != [(x < 215 and y < 237 and z < 225)]:
        triggered.add(28)

    if [((x + y / 10 + z) > 288 and (x + y / 10 + z) < 308)] != [
        ((x + y / 10 + z) > 288 and (x + y / 10 + z) < 508)]:
        triggered.add(29)
    if [((x + y / 10 + z) > 288 and (x + y / 10 + z) < 308)] != [
        ((x + y / 10 + z) > 288 and (x + y / 10 + z) < 708)]:
        triggered.add(30)

    if [(x * y / 100 > 440 and x * y / 100 < 500)] != [
        (x * y / 100 > 440 and x * y / 100 < 900)]:
        triggered.add(31)
    if [(x * y / 100 > 440 and x * y / 100 < 500)] != [
        (x * y / 100 > 440 and x * y / 100 < 440)]:
        triggered.add(32)

    if [(x * z > 12500 and x * z < 14500)] != [(x * z > 12500 and x * z < 17600)]:
        triggered.add(33)
    if [(x * z > 12500 and x * z < 14500)] != [(x * z > 12500 and x * z < 54500)]:
        triggered.add(34)

    if [(y / 10 * z > 1400 and y / 10 * z < 1600)] != [
        (y / 10 * z > 1400 and y / 10 * z < 7800)]:
        triggered.add(35)
    if [(y / 10 * z > 1400 and y / 10 * z < 1600)] != [
        (y / 10 * z > 1400 and y / 10 * z < 5600)]:
        triggered.add(36)

    if [((x + y / 10 + z) / 3 > 96 and (x + y / 10 + z) / 3 < 102)] != [
        ((x + y / 10 + z) / 3 > 96 and (x + y / 10 + z) / 3 < 1052)]:
        triggered.add(37)
    if [((x + y / 10 + z) / 3 > 96 and (x + y / 10 + z) / 3 < 102)] != [
        ((x + y / 10 + z) / 3 > 96 and (x + y / 10 + z) / 3 < 1502)]:
        triggered.add(38)

    if [(math.sqrt(x ** 2 + (y / 10) ** 2 + z ** 2) > 220)] != [
        (math.sqrt(x ** 2 + (y / 10) ** 2 + z ** 2) > 2520)]:
        triggered.add(39)
    if [(math.sqrt(x ** 2 + (y / 10) ** 2 + z ** 2) > 220)] != [
        (math.sqrt(x ** 2 + (y / 10) ** 2 + z ** 2) > 280)]:
        triggered.add(40)

    if [(x / (y / 10) > 8.2 and x / (y / 10) < 9.2)] != [
        (x / (y / 10) > 8.2 and x / (y / 10) < 92)]:
        triggered.add(41)
    if [(x / (y / 10) > 8.2 and x / (y / 10) < 9.2)] != [
        (x / (y / 10) > 8.2 and x / (y / 10) < 19.2)]:
        triggered.add(42)

    if [(x / z > 2.8 and x / z < 3.4)] != [(x / z > 2.8 and x / z < 13.4)]:
        triggered.add(43)
    if [(x / z > 2.8 and x / z < 3.4)] != [(x / z > 2.8 and x / z < 31.4)]:
        triggered.add(44)

    if [((y / 10) / z > 0.32 and (y / 10) / z < 0.38)] != [
        ((y / 10) / z > 0.32 and (y / 10) / z < 38)]:
        triggered.add(45)
    if [((y / 10) / z > 0.32 and (y / 10) / z < 0.38)] != [
        ((y / 10) / z > 0.32 and (y / 10) / z < 3.8)]:
        triggered.add(46)

    if [((x - 180) + (y - 220) / 10 > 22 and (x - 180) + (y - 220) / 10 < 32)] != [
        ((x - 180) + (y - 220) / 10 > 22 and (x - 180) + (y - 220) / 10 < 312)]:
        triggered.add(47)
    if [((x - 180) + (y - 220) / 10 > 22 and (x - 180) + (y - 220) / 10 < 32)] != [
        ((x - 180) + (y - 220) / 10 > 22 and (x - 180) + (y - 220) / 10 < 132)]:
        triggered.add(48)

    if [((x - 180) + (z - 50) > 35 and (x - 180) + (z - 50) < 45)] != [
        ((x - 180) + (z - 50) > 35 and (x - 180) + (z - 50) < 415)]:
        triggered.add(49)
    if [((x - 180) + (z - 50) > 35 and (x - 180) + (z - 50) < 45)] != [
        ((x - 180) + (z - 50) > 35 and (x - 180) + (z - 50) < 145)]:
        triggered.add(50)

    if [((y - 220) / 10 + (z - 50) > 25 and (y - 220) / 10 + (z - 50) < 35)] != [
        ((y - 220) / 10 + (z - 50) > 25 and (y - 220) / 10 + (z - 50) < 315)]:
        triggered.add(51)
    if [((y - 220) / 10 + (z - 50) > 25 and (y - 220) / 10 + (z - 50) < 35)] != [
        ((y - 220) / 10 + (z - 50) > 25 and (y - 220) / 10 + (z - 50) < 135)]:
        triggered.add(52)

    if [(abs((x - 200) - (y - 230) / 2) < 8)] != [(abs((x - 200) - (y - 230) / 2) < 18)]:
        triggered.add(53)
    if [(abs((x - 200) - (y - 230) / 2) < 8)] != [(abs((x - 200) - (y - 230) / 2) < 81)]:
        triggered.add(54)

    if [(abs((x - 200) - (z - 65) * 3) < 12)] != [(abs((x - 200) - (z - 65) * 3) < 92)]:
        triggered.add(55)
    if [(abs((x - 200) - (z - 65) * 3) < 12)] != [(abs((x - 200) - (z - 65) * 3) < 122)]:
        triggered.add(56)

    if [(abs((y - 230) / 10 - (z - 65) / 5) < 3)] != [(abs((y - 230) / 10 - (z - 65) / 5) < 13)]:
        triggered.add(57)
    if [(abs((y - 230) / 10 - (z - 65) / 5) < 3)] != [(abs((y - 230) / 10 - (z - 65) / 5) < 31)]:
        triggered.add(58)

    if [(x / (y / 10 + 5) > 6.5 and x / (y / 10 + 5) < 8.5)] != [
        (x / (y / 10 + 5) > 6.5 and x / (y / 10 + 5) < 85)]:
        triggered.add(59)
    if [(x / (y / 10 + 5) > 6.5 and x / (y / 10 + 5) < 8.5)] != [
        (x / (y / 10 + 5) > 6.5 and x / (y / 10 + 5) < 81.5)]:
        triggered.add(60)

    if [((y / 10) / (z + 15) > 0.26 and (y / 10) / (z + 15) < 0.34)] != [
        ((y / 10) / (z + 15) > 0.26 and (y / 10) / (z + 15) < 10.34)]:
        triggered.add(61)
    if [((y / 10) / (z + 15) > 0.26 and (y / 10) / (z + 15) < 0.34)] != [
        ((y / 10) / (z + 15) > 0.26 and (y / 10) / (z + 15) < 40.34)]:
        triggered.add(62)

    if [(z / (x / 5) > 1.4 and z / (x / 5) < 1.8)] != [(z / (x / 5) > 1.4 and z / (x / 5) < 11.8)]:
        triggered.add(63)
    if [(z / (x / 5) > 1.4 and z / (x / 5) < 1.8)] != [(z / (x / 5) > 1.4 and z / (x / 5) < 41.8)]:
        triggered.add(64)

    if [((x * y * z / 10000) > 28 and (x * y * z / 10000) < 36)] != [
        ((x * y * z / 10000) > 28 and (x * y * z / 10000) < 332)]:
        triggered.add(65)
    if [((x * y * z / 10000) > 28 and (x * y * z / 10000) < 36)] != [
        ((x * y * z / 10000) > 28 and (x * y * z / 10000) < 236)]:
        triggered.add(66)

    if [(x * 0.5 + y / 10 * 0.3 + z * 0.2 > 125 and x * 0.5 + y / 10 * 0.3 + z * 0.2 < 135)] != [
        (x * 0.5 + y / 10 * 0.3 + z * 0.2 > 125 and x * 0.5 + y / 10 * 0.3 + z * 0.2 < 935)]:
        triggered.add(67)
    if [(x * 0.5 + y / 10 * 0.3 + z * 0.2 > 125 and x * 0.5 + y / 10 * 0.3 + z * 0.2 < 135)] != [
        (x * 0.5 + y / 10 * 0.3 + z * 0.2 > 125 and x * 0.5 + y / 10 * 0.3 + z * 0.2 < 635)]:
        triggered.add(68)

    if [((x / 200) ** 0.4 * (y / 230) ** 0.3 * (z / 65) ** 0.3 > 0.92)] != [
        ((x / 200) ** 0.4 * (y / 230) ** 0.3 * (z / 65) ** 0.3 > 9.2)]:
        triggered.add(69)
    if [((x / 200) ** 0.4 * (y / 230) ** 0.3 * (z / 65) ** 0.3 > 0.92)] != [
        ((x / 200) ** 0.4 * (y / 230) ** 0.3 * (z / 65) ** 0.3 > 92)]:
        triggered.add(70)

    if [((x - 200) * (y - 230) / 10 > -80 and (x - 200) * (y - 230) / 10 < 80)] != [
        ((x - 200) * (y - 230) / 10 > -80 and (x - 200) * (y - 230) / 10 < 180)]:
        triggered.add(71)
    if [((x - 200) * (y - 230) / 10 > -80 and (x - 200) * (y - 230) / 10 < 80)] != [
        ((x - 200) * (y - 230) / 10 > -80 and (x - 200) * (y - 230) / 10 < 810)]:
        triggered.add(72)

    if [((x - 200) * (z - 65) > -120 and (x - 200) * (z - 65) < 120)] != [
        ((x - 200) * (z - 65) > -120 and (x - 200) * (z - 65) < 820)]:
        triggered.add(73)
    if [((x - 200) * (z - 65) > -120 and (x - 200) * (z - 65) < 120)] != [
        ((x - 200) * (z - 65) > -120 and (x - 200) * (z - 65) < 520)]:
        triggered.add(74)

    if [((y - 230) / 10 * (z - 65) > -30 and (y - 230) / 10 * (z - 65) < 30)] != [
        ((y - 230) / 10 * (z - 65) > -30 and (y - 230) / 10 * (z - 65) < 310)]:
        triggered.add(75)
    if [((y - 230) / 10 * (z - 65) > -30 and (y - 230) / 10 * (z - 65) < 30)] != [
        ((y - 230) / 10 * (z - 65) > -30 and (y - 230) / 10 * (z - 65) < 130)]:
        triggered.add(76)

    if [(abs(x / 200 + y / 230 + z / 65 - 3) < 0.18)] != [
        (abs(x / 200 + y / 230 + z / 65 - 3) < 18)]:
        triggered.add(77)
    if [(abs(x / 200 + y / 230 + z / 65 - 3) < 0.18)] != [
        (abs(x / 200 + y / 230 + z / 65 - 3) < 10.18)]:
        triggered.add(78)

    if [(abs((x / 200) * (y / 230) * (z / 65) - 1) < 0.12)] != [
        (abs((x / 200) * (y / 230) * (z / 65) - 1) < 10.12)]:
        triggered.add(79)
    if [(abs((x / 200) * (y / 230) * (z / 65) - 1) < 0.12)] != [
        (abs((x / 200) * (y / 230) * (z / 65) - 1) < 12.12)]:
        triggered.add(80)

    if [(x > 197 and x < 203 and y > 228 and y < 232 and z > 63.5 and z < 66.5)] != [
        (x > 197 and x < 203 and y > 228 and y < 232 and z > 63.5 and z < 616.5)]:
        triggered.add(81)
    if [(x > 197 and x < 203 and y > 228 and y < 232 and z > 63.5 and z < 66.5)] != [
        (x > 197 and x < 203 and y > 228 and y < 232 and z > 63.5 and z < 166.5)]:
        triggered.add(82)

    if [(abs(x / (y / 10) - 8.7) < 0.3 and abs(z - 65) < 1.5)] != [
        (abs(x / (y / 10) - 8.7) < 0.3 and abs(z - 65) < 11.5)]:
        triggered.add(83)
    if [(abs(x / (y / 10) - 8.7) < 0.3 and abs(z - 65) < 1.5)] != [
        (abs(x / (y / 10) - 8.7) < 0.3 and abs(z - 65) < 15)]:
        triggered.add(84)

    if [(x * y / 100 > 455 and x * y / 100 < 465 and z > 64 and z < 66)] != [
        (x * y / 100 > 455 and x * y / 100 < 465 and z > 64 and z < 166)]:
        triggered.add(85)
    if [(x * y / 100 > 455 and x * y / 100 < 465 and z > 64 and z < 66)] != [
        (x * y / 100 > 455 and x * y / 100 < 465 and z > 64 and z < 616)]:
        triggered.add(86)

    if [((x + y / 10) / 2 > 112 and (x + y / 10) / 2 < 116 and z > 63.5 and z < 66.5)] != [
        ((x + y / 10) / 2 > 112 and (x + y / 10) / 2 < 116 and z > 63.5 and z < 166.5)]:
        triggered.add(87)
    if [((x + y / 10) / 2 > 112 and (x + y / 10) / 2 < 116 and z > 63.5 and z < 66.5)] != [
        ((x + y / 10) / 2 > 112 and (x + y / 10) / 2 < 116 and z > 63.5 and z < 661.5)]:
        triggered.add(88)

    if [(abs(x - y / 10) < 175 and z > 63.5 and z < 66.5)] != [(abs(x - y / 10) < 175 and z > 63.5 and z < 616.5)]:
        triggered.add(89)
    if [(abs(x - y / 10) < 175 and z > 63.5 and z < 66.5)] != [(abs(x - y / 10) < 175 and z > 63.5 and z < 6615)]:
        triggered.add(90)

    if [(math.sqrt((x - 200) ** 2 + (y / 10 - 23) ** 2) < 3 and abs(z - 65) < 1.5)] != [
        (math.sqrt((x - 200) ** 2 + (y / 10 - 23) ** 2) < 3 and abs(z - 65) < 15)]:
        triggered.add(91)
    if [(math.sqrt((x - 200) ** 2 + (y / 10 - 23) ** 2) < 3 and abs(z - 65) < 1.5)] != [
        (math.sqrt((x - 200) ** 2 + (y / 10 - 23) ** 2) < 3 and abs(z - 65) < 11.5)]:
        triggered.add(92)

    if [(x / 200 > 0.985 and x / 200 < 1.015 and y / 230 > 0.987 and y / 230 < 1.013)] != [
        (x / 200 > 0.985 and x / 200 < 1.015 and y / 230 > 0.987 and y / 230 < 10.13)]:
        triggered.add(93)
    if [(x / 200 > 0.985 and x / 200 < 1.015 and y / 230 > 0.987 and y / 230 < 1.013)] != [
        (x / 200 > 0.985 and x / 200 < 1.015 and y / 230 > 0.987 and y / 230 < 1013)]:
        triggered.add(94)

    if [(z / 65 > 0.985 and z / 65 < 1.015)] != [(z / 65 > 0.985 and z / 65 < 121.015)]:
        triggered.add(95)
    if [(z / 65 > 0.985 and z / 65 < 1.015)] != [(z / 65 > 0.985 and z / 65 < 111.015)]:
        triggered.add(96)

    if [((x / 200 + y / 230 + z / 65) / 3 > 0.985 and (x / 200 + y / 230 + z / 65) / 3 < 1.015)] != [
        ((x / 200 + y / 230 + z / 65) / 3 > 0.985 and (x / 200 + y / 230 + z / 65) / 3 < 2.115)]:
        triggered.add(97)
    if [((x / 200 + y / 230 + z / 65) / 3 > 0.985 and (x / 200 + y / 230 + z / 65) / 3 < 1.015)] != [
        ((x / 200 + y / 230 + z / 65) / 3 > 0.985 and (x / 200 + y / 230 + z / 65) / 3 < 15)]:
        triggered.add(98)

    if [(max(abs(x / 200 - 1), abs(y / 230 - 1), abs(z / 65 - 1)) < 0.02)] != [
        (max(abs(x / 200 - 1), abs(y / 230 - 1), abs(z / 65 - 1)) < 22)]:
        triggered.add(99)
    if [(max(abs(x / 200 - 1), abs(y / 230 - 1), abs(z / 65 - 1)) < 0.02)] != [
        (max(abs(x / 200 - 1), abs(y / 230 - 1), abs(z / 65 - 1)) < 10.02)]:
        triggered.add(100)

    if [(min(x / 200, y / 230, z / 65) > 0.985)] != [(min(x / 200, y / 230, z / 65) > 9.85)]:
        triggered.add(101)
    if [(min(x / 200, y / 230, z / 65) > 0.985)] != [(min(x / 200, y / 230, z / 65) > 10.985)]:
        triggered.add(102)

    if [(max(x / 200, y / 230, z / 65) < 1.015)] != [(max(x / 200, y / 230, z / 65) < 10.15)]:
        triggered.add(103)
    if [(max(x / 200, y / 230, z / 65) < 1.015)] != [(max(x / 200, y / 230, z / 65) < 15)]:
        triggered.add(104)

    if [(abs(max(x, y / 10 * 8.7, z) - min(x, y / 10 * 8.7, z)) < 15)] != [
        (abs(max(x, y / 10 * 8.7, z) - min(x, y / 10 * 8.7, z)) < 125)]:
        triggered.add(105)
    if [(abs(max(x, y / 10 * 8.7, z) - min(x, y / 10 * 8.7, z)) < 15)] != [
        (abs(max(x, y / 10 * 8.7, z) - min(x, y / 10 * 8.7, z)) < 88)]:
        triggered.add(106)

    if [((x + y / 10 + z) > 297 and (x + y / 10 + z) < 299)] != [
        ((x + y / 10 + z) > 297 and (x + y / 10 + z) < 99)]:
        triggered.add(107)
    if [((x + y / 10 + z) > 297 and (x + y / 10 + z) < 299)] != [
        ((x + y / 10 + z) > 297 and (x + y / 10 + z) < 549)]:
        triggered.add(108)

    if [((x * y * z / 10000) > 31.5 and (x * y * z / 10000) < 345)] != [
        ((x * y * z / 10000) > 31.5 and (x * y * z / 10000) < 132.5)]:
        triggered.add(109)
    if [((x * y * z / 10000) > 31.5 and (x * y * z / 10000) < 32.5)] != [
        ((x * y * z / 10000) > 31.5 and (x * y * z / 10000) < 232.5)]:
        triggered.add(110)

    if [(abs((x + y / 10 + z) / 3 - 99.3) < 0.8)] != [(abs((x + y / 10 + z) / 3 - 99.3) < 8)]:
        triggered.add(111)
    if [(abs((x + y / 10 + z) / 3 - 99.3) < 0.8)] != [(abs((x + y / 10 + z) / 3 - 99.3) < 10.8)]:
        triggered.add(112)

    if [(x > 199 and x < 201 and y > 229 and y < 231)] != [(x > 199 and x < 201 and y > 229 and y < 2231)]:
        triggered.add(113)
    if [(x > 199 and x < 201 and y > 229 and y < 231)] != [(x > 199 and x < 201 and y > 229 and y < 1231)]:
        triggered.add(114)

    if [(z > 64.5 and z < 65.5)] != [(z > 64.5 and z < 615.5)]:
        triggered.add(115)
    if [(z > 64.5 and z < 65.5)] != [(z > 64.5 and z < 655)]:
        triggered.add(116)

    if [((x - 200) ** 2 + (y / 10 - 23) ** 2 + (z - 65) ** 2) < 9] != [
        ((x - 200) ** 2 + (y / 10 - 23) ** 2 + (z - 65) ** 2) < 19]:
        triggered.add(117)
    if [((x - 200) ** 2 + (y / 10 - 23) ** 2 + (z - 65) ** 2) < 9] != [
        ((x - 200) ** 2 + (y / 10 - 23) ** 2 + (z - 65) ** 2) < 91]:
        triggered.add(118)

    if [(abs(x / (y / 10) / z - 0.134) < 0.005)] != [(abs(x / (y / 10) / z - 0.134) < 10.005)]:
        triggered.add(119)
    if [(abs(x / (y / 10) / z - 0.134) < 0.005)] != [(abs(x / (y / 10) / z - 0.134) < 5)]:
        triggered.add(120)

    if [(x < 188 or y < 225 or z < 58)] != [(x < 188 or y < 225 or z < 158)]:
        triggered.add(121)
    if [(x < 188 or y < 225 or z < 58)] != [(x < 188 or y < 225 or 28 < 58)]:
        triggered.add(122)

    if [(x > 212 or y > 235 or z > 72)] != [(x > 212 or y > 235 or 12 * z > 72)]:
        triggered.add(123)
    if [(x > 212 or y > 235 or z > 72)] != [(x > 212 or y > 235 or 88 > 72)]:
        triggered.add(124)

    if [(abs(x / (y / 10) - 8.7) > 0.5)] != [(abs(x / (y / 10) - 8.7) > 10.5)]:
        triggered.add(125)
    if [(abs(x / (y / 10) - 8.7) > 0.5)] != [(abs(x / (y / 10) - 8.7) > 5)]:
        triggered.add(126)

    if [(abs(x / z - 3.08) > 0.3)] != [(abs(x / z - 3.08) > 10.3)]:
        triggered.add(127)
    if [(abs(x / z - 3.08) > 0.3)] != [(abs(x / z - 3.08) > 123)]:
        triggered.add(128)

    if [(abs((y / 10) / z - 0.354) > 0.03)] != [(abs((y / 10) / z - 0.354) > 3)]:
        triggered.add(129)
    if [(abs((y / 10) / z - 0.354) > 0.03)] != [(abs((y / 10) / z - 0.354) > 10.03)]:
        triggered.add(130)

    if [((x + y / 10 + z) < 292 or (x + y / 10 + z) > 304)] != [
        ((x + y / 10 + z) < 292 or (x + y / 10 + z) > 3224)]:
        triggered.add(131)
    if [((x + y / 10 + z) < 292 or (x + y / 10 + z) > 304)] != [
        ((x + y / 10 + z) < 292 or (x + y / 10 + z) > 1304)]:
        triggered.add(132)

    if [((x * y * z / 10000) < 29 or (x * y * z / 10000) > 35)] != [
        ((x * y * z / 10000) < 29 or (x * y * z / 10000) > 135)]:
        triggered.add(133)
    if [((x * y * z / 10000) < 29 or (x * y * z / 10000) > 35)] != [
        ((x * y * z / 10000) < 29 or (x * y * z / 10000) > 315)]:
        triggered.add(134)

    if [(x > 215)] != [(x > 1215)]:
        triggered.add(135)
    if [(x > 215)] != [(x > 2115)]:
        triggered.add(136)

    if [(x < 185)] != [(x < 1815)]:
        triggered.add(137)
    if [(x < 185)] != [(x < 1185)]:
        triggered.add(138)

    if [(y > 237)] != [(y > 2137)]:
        triggered.add(139)
    if [(y > 237)] != [(y > 2317)]:
        triggered.add(140)

    if [(y < 223)] != [(y < 2123)]:
        triggered.add(141)
    if [(y < 223)] != [(y < 2283)]:
        triggered.add(142)

    if [(z > 75)] != [(z > 751)]:
        triggered.add(143)
    if [(z > 75)] != [(z > 175)]:
        triggered.add(144)

    if [(z < 55)] != [(z < 515)]:
        triggered.add(145)
    if [(z < 55)] != [(z < 559)]:
        triggered.add(146)

    if [(x > 215 or x < 185) and (y > 237 or y < 223)] != [(x > 215 or x < 185) and (y > 237 or y < 2213)]:
        triggered.add(147)
    if [(x > 215 or x < 185) and (y > 237 or y < 223)] != [(x > 215 or x < 185) and (y > 237 or 200 < 223)]:
        triggered.add(148)

    if [(x > 215 or x < 185) and (z > 75 or z < 55)] != [(x > 215 or x < 185) and (z > 175 or z < 55)]:
        triggered.add(149)
    if [(x > 215 or x < 185) and (z > 75 or z < 55)] != [(x > 215 or x < 185) and (z > 75 or z < 551)]:
        triggered.add(150)

    if [(y > 237 or y < 223) and (z > 75 or z < 55)] != [(y > 237 or y < 223) and (z > 75 or z < 515)]:
        triggered.add(151)
    if [(y > 237 or y < 223) and (z > 75 or z < 55)] != [(y > 237 or y < 223) and (z > 175 or z < 55)]:
        triggered.add(152)

    if [(x > 215 or x < 185) and (y > 237 or y < 223) and (z > 75 or z < 55)] != [
        (x > 215 or x < 1185) and (y > 237 or y < 223) and (z > 75 or z < 55)]:
        triggered.add(153)
    if [(x > 215 or x < 185) and (y > 237 or y < 223) and (z > 75 or z < 55)] != [
        (x > 215 or x < 185) and (y > 2317 or y < 223) and (z > 75 or z < 55)]:
        triggered.add(154)

    if [(abs(x - 200) > 20) or (abs(y - 230) > 15) or (abs(z - 65) > 10)] != [
        (abs(x - 200) > 20) or (abs(y - 230) > 115) or (abs(z - 65) > 10)]:
        triggered.add(155)
    if [(abs(x - 200) > 20) or (abs(y - 230) > 15) or (abs(z - 65) > 10)] != [
        (abs(x - 200) > 20) or (abs(y - 230) > 415) or (abs(z - 65) > 10)]:
        triggered.add(156)

    if [(x > 220) or (y > 240) or (z > 80)] != [(x > 220) or (y > 240) or (z > 810)]:
        triggered.add(157)
    if [(x > 220) or (y > 240) or (z > 80)] != [(x > 220) or (y > 240) or (z > 911)]:
        triggered.add(158)
    if [(x < 175) or (y < 215) or (z < 45)] != [(x < 175) or (y < 215) or (z < 415)]:
        triggered.add(159)
    if [(x < 175) or (y < 215) or (z < 45)] != [(x < 175) or (y < 215) or (z < 145)]:
        triggered.add(160)
    return triggered

# === 归一化/反归一化函数（已适配新范围） ===
def normalize_state(state):
    """将原始状态映射到[0,1]"""
    x, y, z = state
    normalized_x = (x - MIN_X) / (MAX_X - MIN_X)
    normalized_y = (y - MIN_Y) / (MAX_Y - MIN_Y)
    normalized_z = (z - MIN_Z) / (MAX_Z - MIN_Z)
    return [normalized_x, normalized_y, normalized_z]

def denormalize_state(normalized_state):
    """将归一化状态还原为原始整数（并裁剪）"""
    norm_x, norm_y, norm_z = normalized_state
    x = int(round(norm_x * (MAX_X - MIN_X) + MIN_X))
    y = int(round(norm_y * (MAX_Y - MIN_Y) + MIN_Y))
    z = int(round(norm_z * (MAX_Z - MIN_Z) + MIN_Z))
    # 裁剪确保在范围内
    x = max(MIN_X, min(MAX_X, x))
    y = max(MIN_Y, min(MAX_Y, y))
    z = max(MIN_Z, min(MAX_Z, z))
    return [x, y, z]

# === reward function ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward

# === 目标路径组 ===
targetPaths = [
    {1, 2, 3, 4, 5, 6, 11, 12, 14, 16, 17, 18, 19, 21, 22, 23, 25, 26, 29, 30, 31, 34, 35, 36, 37, 38, 39, 40, 41, 42,
     47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100,
     101, 102, 103, 104, 108, 112, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137,
     138, 141, 142, 143, 144, 145, 146, 147, 148, 149, 157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 11, 12, 14, 15, 17, 18, 19, 23, 25, 26, 29, 30, 32, 34, 35, 36, 37, 38, 39, 40, 41, 42, 47, 48,
     49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102,
     103, 104, 108, 112, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 141,
     142, 143, 144, 145, 146, 147, 148, 149, 157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 11, 12, 14, 16, 17, 18, 19, 23, 25, 26, 29, 30, 31, 34, 35, 36, 37, 38, 39, 40, 41, 42, 47, 48,
     49, 50, 53, 54, 55, 56, 57, 58, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102, 103,
     104, 108, 111, 112, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 141,
     142, 143, 144, 145, 146, 147, 148, 149, 157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 14, 16, 17, 18, 19, 22, 25, 26, 29, 30, 31, 34, 35, 36, 37, 38, 39, 40, 41, 42, 47, 48, 49, 50,
     51, 52, 54, 55, 56, 57, 58, 63, 64, 67, 68, 69, 70, 73, 75, 76, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102,
     103, 104, 108, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 139, 140,
     141, 142, 143, 144, 145, 146, 149, 152, 154, 159, 160},

    {1, 2, 3, 4, 5, 6, 14, 16, 17, 18, 19, 21, 22, 23, 25, 26, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 47,
     48, 49, 50, 53, 54, 55, 56, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104,
     108, 111, 112, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 139, 140,
     141, 142, 143, 144, 145, 146, 149, 152, 154, 159, 160},

    {1, 2, 3, 4, 5, 6, 9, 10, 11, 12, 14, 17, 18, 19, 20, 21, 22, 25, 26, 27, 28, 29, 30, 32, 34, 35, 36, 37, 38, 39,
     40, 41, 42, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 85, 86, 87, 88, 95, 96,
     97, 98, 99, 100, 103, 104, 105, 108, 111, 112, 115, 116, 119, 120, 125, 126, 127, 128, 129, 130, 131, 132, 137,
     138, 141, 142, 143, 144, 145, 146, 157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 11, 12, 14, 17, 18, 19, 21, 22, 23, 25, 26, 29, 30, 31, 34, 35, 36, 37, 38, 39, 40, 41, 42, 47,
     48, 49, 50, 51, 52, 54, 55, 56, 57, 58, 59, 60, 63, 64, 67, 68, 69, 70, 73, 77, 78, 79, 80, 95, 96, 97, 98, 99,
     100, 103, 104, 108, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 141,
     142, 143, 144, 145, 146, 147, 148, 149, 159, 160},

    {1, 2, 3, 4, 5, 6, 9, 10, 14, 16, 17, 18, 19, 20, 21, 22, 25, 26, 29, 30, 31, 34, 35, 36, 37, 38, 39, 40, 47, 48,
     49, 50, 51, 52, 55, 56, 57, 58, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102, 103,
     104, 105, 108, 111, 112, 115, 116, 119, 120, 121, 122, 127, 128, 129, 130, 131, 132, 137, 138, 139, 140, 141, 142,
     143, 144, 145, 146, 152, 153, 157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 13, 16, 17, 18, 21, 22, 25, 26, 29, 30, 32, 34, 35, 36, 37, 38, 39, 40, 49, 50, 51,
     52, 55, 56, 57, 58, 63, 64, 69, 70, 77, 78, 79, 80, 87, 88, 93, 94, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104,
     105, 108, 111, 112, 115, 116, 119, 120, 121, 122, 127, 128, 129, 130, 131, 132, 137, 138, 139, 140, 141, 142, 143,
     144, 145, 146, 152, 153, 157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 16, 17, 18, 19, 20, 21, 22, 25, 26, 29, 30, 32, 33, 34, 35, 36, 37, 38, 49, 50, 51,
     52, 53, 54, 55, 56, 57, 58, 63, 64, 69, 70, 77, 78, 79, 80, 89, 90, 95, 96, 97, 98, 99, 100, 103, 104, 105, 108,
     111, 112, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 137, 138, 139, 140, 141, 142, 143,
     144, 145, 146, 152, 153, 157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 14, 16, 17, 18, 19, 21, 22, 23, 25, 26, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40, 47, 48, 49,
     50, 53, 54, 55, 56, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 84, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 108,
     111, 112, 115, 116, 119, 120, 121, 122, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 139, 140, 141, 142, 143,
     144, 145, 146, 149, 152, 154, 159, 160},

    {1, 2, 3, 4, 5, 6, 11, 12, 14, 16, 17, 18, 19, 21, 22, 23, 24, 25, 26, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40,
     41, 42, 47, 48, 49, 50, 54, 55, 56, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104,
     108, 111, 112, 115, 116, 119, 120, 121, 122, 125, 126, 129, 130, 131, 132, 135, 136, 137, 138, 141, 142, 143, 144,
     145, 146, 147, 148, 149, 159, 160},

    {1, 2, 3, 4, 5, 6, 14, 16, 17, 18, 19, 21, 22, 23, 24, 25, 26, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42,
     43, 44, 47, 48, 49, 50, 54, 55, 56, 67, 68, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102, 103,
     104, 108, 112, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 131, 132, 135, 136, 137, 138, 139, 140, 141, 142,
     145, 146, 150, 151, 159, 160},

    {1, 2, 3, 4, 9, 10, 11, 12, 13, 15, 17, 18, 25, 26, 27, 28, 29, 30, 32, 34, 35, 36, 37, 38, 39, 40, 49, 50, 51, 52,
     55, 56, 57, 58, 63, 64, 69, 70, 77, 78, 79, 80, 81, 82, 85, 86, 87, 88, 95, 96, 97, 98, 99, 100, 101, 102, 103,
     104, 105, 108, 111, 112, 115, 116, 119, 120, 121, 122, 127, 128, 129, 130, 131, 132, 137, 138, 141, 142, 143, 144,
     145, 146, 157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 14, 16, 17, 18, 19, 22, 23, 24, 25, 26, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40, 47, 48, 49,
     50, 53, 54, 55, 56, 69, 70, 73, 74, 77, 78, 79, 80, 83, 84, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 108, 111,
     112, 115, 116, 119, 120, 121, 122, 129, 130, 131, 132, 135, 136, 137, 138, 139, 140, 141, 142, 143, 144, 145, 146,
     149, 152, 154, 159, 160},

    {1, 2, 3, 4, 5, 6, 9, 10, 11, 12, 13, 15, 17, 18, 23, 25, 26, 27, 28, 32, 33, 34, 35, 36, 37, 38, 49, 50, 55, 56,
     57, 58, 63, 64, 69, 70, 77, 78, 79, 80, 81, 82, 85, 86, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 105, 108, 111,
     112, 113, 114, 115, 116, 119, 120, 121, 122, 127, 128, 129, 130, 131, 132, 137, 138, 141, 142, 143, 144, 145, 146,
     157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 13, 16, 17, 18, 22, 23, 24, 25, 26, 32, 33, 34, 35, 36, 49, 50, 55, 56, 63, 64, 69,
     70, 77, 78, 79, 80, 87, 88, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 108, 111, 112, 113, 114,
     115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 137, 138, 139, 140, 141, 142, 143, 144, 145, 146, 152,
     153, 159, 160},

    {1, 2, 3, 4, 5, 6, 14, 16, 19, 22, 23, 24, 25, 26, 29, 30, 31, 33, 34, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47,
     48, 49, 50, 54, 55, 56, 67, 68, 69, 70, 77, 78, 79, 80, 97, 98, 99, 100, 103, 104, 108, 111, 112, 119, 120, 121,
     122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 139, 140, 141, 142, 145, 146, 150, 151, 159, 160},

    {1, 2, 5, 6, 14, 16, 17, 18, 19, 22, 23, 24, 25, 26, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40, 47, 48, 49, 50, 53,
     54, 69, 70, 73, 74, 77, 78, 79, 80, 83, 84, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 108, 111, 112, 115, 116,
     119, 120, 121, 122, 131, 132, 135, 136, 137, 138, 139, 140, 141, 142, 145, 146, 150, 151, 155, 156, 159, 160},

    {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 13, 16, 17, 18, 22, 23, 24, 25, 26, 32, 33, 34, 35, 36, 53, 54, 55, 56, 63, 64, 69,
     70, 77, 78, 79, 80, 87, 88, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 107, 113, 114, 115, 116,
     118, 119, 120, 121, 122, 125, 126, 127, 128, 137, 138, 139, 140, 141, 142, 145, 146, 151, 155, 156, 159, 160},

    {1, 2, 5, 6, 9, 10, 14, 17, 18, 19, 20, 22, 23, 27, 28, 29, 30, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 49, 50, 54,
     55, 56, 57, 58, 63, 64, 69, 70, 73, 74, 79, 80, 87, 88, 95, 96, 97, 98, 99, 100, 103, 104, 108, 109, 111, 112, 115,
     116, 119, 120, 125, 126, 127, 128, 129, 130, 131, 132, 137, 138, 143, 144, 145, 146, 152, 153, 157, 158},

    {1, 2, 3, 4, 5, 6, 14, 16, 19, 22, 23, 24, 25, 26, 29, 30, 31, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49,
     50, 54, 55, 56, 57, 58, 61, 62, 65, 67, 68, 69, 70, 97, 98, 99, 100, 103, 104, 108, 109, 111, 112, 119, 120, 125,
     126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 139, 140, 141, 142, 145, 146, 150, 151, 159, 160},

    {1, 2, 3, 4, 5, 6, 14, 16, 19, 21, 22, 23, 24, 25, 26, 29, 30, 31, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48,
     49, 50, 54, 55, 56, 65, 69, 70, 97, 98, 99, 100, 103, 104, 108, 109, 111, 112, 119, 120, 121, 122, 125, 126, 127,
     128, 129, 130, 131, 132, 134, 135, 136, 137, 138, 141, 142, 145, 146, 147, 148, 150, 159, 160},

    {1, 2, 3, 4, 5, 6, 7, 8, 13, 16, 17, 18, 22, 23, 24, 25, 26, 32, 35, 36, 53, 54, 69, 70, 79, 80, 87, 88, 91, 92, 93,
     94, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 109, 111, 112, 115, 116, 117, 118, 119, 120, 121, 122, 125, 126,
     137, 138, 139, 140, 141, 142, 145, 146, 151, 155, 156, 159, 160},

    {1, 2, 3, 4, 5, 6, 9, 10, 11, 12, 14, 15, 17, 18, 19, 20, 23, 24, 25, 26, 32, 33, 34, 35, 36, 39, 40, 49, 50, 53,
     54, 69, 70, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 108, 109, 111, 112, 115, 116, 119, 120, 121, 122,
     123, 124, 129, 130, 137, 138, 141, 142, 145, 146, 159, 160},

    {3, 4, 5, 6, 14, 16, 19, 21, 22, 31, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 54, 56, 57, 58, 61, 62, 65, 66,
     67, 68, 69, 70, 79, 80, 99, 100, 103, 104, 108, 109, 110, 111, 112, 119, 120, 125, 126, 127, 128, 129, 130, 131,
     132, 134, 135, 136, 137, 138, 141, 142, 147, 148},

    {1, 2, 5, 6, 17, 18, 19, 22, 27, 28, 33, 34, 35, 36, 51, 52, 54, 56, 57, 58, 63, 64, 65, 69, 70, 71, 72, 89, 90, 95,
     96, 97, 98, 99, 100, 103, 104, 105, 106, 109, 112, 115, 116, 119, 120, 125, 126, 127, 128, 129, 130, 134, 143, 144,
     145, 146, 149, 152, 157, 158},

    {1, 2, 3, 4, 14, 19, 21, 22, 31, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 54, 57, 58, 59, 60, 61, 62, 65, 66, 75, 76,
     77, 78, 79, 80, 99, 100, 103, 104, 110, 111, 112, 119, 120, 125, 126, 127, 128, 129, 130, 133, 134, 135, 136, 137,
     138, 141, 142, 147, 148},
]

target_paths = [set(path) for path in targetPaths]

# === Jaccard相似度 ===
def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0

# === 路径相似度矩阵与分组 ===
def compute_path_similarity_matrix(paths):
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix

def group_paths_by_similarity(paths):
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.mean(avg_sim_scores)
    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]
    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)
    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group

# === 样本生成辅助函数（鲁棒性、Q值等） ===
def compute_robustness(state, path):
    base = section9_hybrid_speed_torque_current(state[0], state[1], state[2])
    if not base:
        return 0.0
    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = [state[0] + dx, state[1] + dy, state[2] + dz]
                neighbor[0] = max(MIN_X, min(MAX_X, neighbor[0]))
                neighbor[1] = max(MIN_Y, min(MAX_Y, neighbor[1]))
                neighbor[2] = max(MIN_Z, min(MAX_Z, neighbor[2]))
                n_trig = section9_hybrid_speed_torque_current(neighbor[0], neighbor[1], neighbor[2])
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0

def compute_q_value_score(state, similar_model):
    """返回 1 - 归一化最大Q值"""
    if similar_model is None:
        return 0.0
    try:
        normalized_state = normalize_state(state)
        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = similar_model(state_tensor)
            max_q_value = torch.max(q_values).item()
            normalized_q = max_q_value / 20.0  # 粗略归一化
            normalized_q = max(0.0, min(1.0, normalized_q))
            return 1.0 - normalized_q
    except:
        return 0.0

# === 生成相似组样本 ===
def generate_samples_for_similar_paths(similar_group, num_candidates=2000, top_k=200, run_id=1):
    SIMILAR_WEIGHTS = [0.55, 0.39, 0.06]

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_similar.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Similar Group Path {path_id} - Run {run_id}\n")
            f.write("weather time_period z\tScore\tSimilarity\tLengthDiff\tRobustness\n")
            for s in samples:
                weather, time_period, z = s['state']
                f.write(
                    f"{weather} {time_period} {z}\t{s['score']:.4f}\t{s['similarity']:.4f}\t{s['length_diff']:.4f}\t{s['robustness']:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_grouped"
    for path_idx in similar_group:
        path = target_paths[path_idx]
        path_id = path_idx + 1
        candidate_samples = []
        attempts = 0
        while len(candidate_samples) < num_candidates and attempts < num_candidates * 10:
            attempts += 1
            weather = np.random.randint(MIN_X, MAX_X + 1)
            time_period = np.random.randint(MIN_Y, MAX_Y + 1)
            z = np.random.randint(MIN_Z, MAX_Z + 1)
            state = (weather, time_period, z)
            triggered = section9_hybrid_speed_torque_current(weather, time_period, z)
            if not triggered:
                continue
            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            candidate_samples.append({
                'state': state,
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            })
        if candidate_samples:
            for sample in candidate_samples:
                score = (SIMILAR_WEIGHTS[0] * sample['similarity'] +
                         SIMILAR_WEIGHTS[1] * sample['length_diff'] +
                         SIMILAR_WEIGHTS[2] * sample['robustness'])
                sample['score'] = score
            candidate_samples.sort(key=lambda x: x['score'], reverse=True)
            selected_samples = candidate_samples[:top_k]
            save_samples(path_id=path_id, samples=selected_samples, base_dir=base_dir)

# === 生成孤立组样本 ===
def generate_samples_for_isolated_paths(isolated_group, similar_model, num_candidates=2000, top_k=200, run_id=1):
    ISOLATED_WEIGHTS = [0.18, 0.21, 0.32, 0.29]

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_isolated.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Isolated Group Path {path_id} - Run {run_id}\n")
            f.write("weather time_period z\tScore\tSimilarity\tLengthDiff\tRobustness\tQValueScore\n")
            for s in samples:
                weather, time_period, z = s['state']
                f.write(
                    f"{weather} {time_period} {z}\t{s['score']:.4f}\t{s['similarity']:.4f}\t{s['length_diff']:.4f}\t{s['robustness']:.4f}\t{s['q_value_score']:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_grouped"
    for path_idx in isolated_group:
        path = target_paths[path_idx]
        path_id = path_idx + 1
        candidate_samples = []
        attempts = 0
        while len(candidate_samples) < num_candidates and attempts < num_candidates * 10:
            attempts += 1
            weather = np.random.randint(MIN_X, MAX_X + 1)
            time_period = np.random.randint(MIN_Y, MAX_Y + 1)
            z = np.random.randint(MIN_Z, MAX_Z + 1)
            state = (weather, time_period, z)
            triggered = section9_hybrid_speed_torque_current(weather, time_period, z)
            if not triggered:
                continue
            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            q_score = compute_q_value_score(state, similar_model)
            candidate_samples.append({
                'state': state,
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'q_value_score': q_score,
                'triggered': triggered
            })
        if candidate_samples:
            for sample in candidate_samples:
                score = (ISOLATED_WEIGHTS[0] * sample['similarity'] +
                         ISOLATED_WEIGHTS[1] * sample['length_diff'] +
                         ISOLATED_WEIGHTS[2] * sample['robustness'] +
                         ISOLATED_WEIGHTS[3] * sample['q_value_score'])
                sample['score'] = score
            candidate_samples.sort(key=lambda x: x['score'], reverse=True)
            selected_samples = candidate_samples[:top_k]
            save_samples(path_id=path_id, samples=selected_samples, base_dir=base_dir)

# === Experience Replay with PER ===
class GroupExperienceReplay:
    def __init__(self, capacity=20000):
        self.capacity = capacity
        self.buffer = deque(maxlen=self.capacity)
        self.priorities = deque(maxlen=self.capacity)

    def append(self, experience):
        self.buffer.append(experience)
        self.priorities.append(experience[-1])

    def sample(self, batch_size, alpha=0.6):
        priorities = np.array(self.priorities) ** alpha
        probabilities = priorities / np.sum(priorities)
        batch_size = min(batch_size, len(self.buffer))
        batch_indices = np.random.choice(len(self.buffer), batch_size, replace=False, p=probabilities)
        batch = [self.buffer[idx] for idx in batch_indices]
        return batch, batch_indices, probabilities[batch_indices]

    def update_priorities(self, batch_indices, td_errors):
        for idx, td_error in zip(batch_indices, td_errors):
            if idx < len(self.priorities):
                self.priorities[idx] = max(td_error, 1e-6)

    def __len__(self):
        return len(self.buffer)

    def get_high_reward_samples(self, target_path, num_samples=20):
        if len(self.buffer) == 0:
            return []
        samples_with_recalculated_scores = []
        seen_states = set()
        for experience in self.buffer:
            state_tensor = experience[0]
            normalized_state = state_tensor.cpu().numpy().flatten()
            state_tuple = tuple(denormalize_state(normalized_state))
            if state_tuple in seen_states:
                continue
            seen_states.add(state_tuple)
            triggered = section9_hybrid_speed_torque_current(state_tuple[0], state_tuple[1], state_tuple[2])
            new_reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_recalculated_scores.append((state_tuple, new_reward, sim, triggered))
        samples_with_recalculated_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_recalculated_scores[:num_samples]

# === 加载样本数据 ===
def load_path_data(file_path):
    path_data = []
    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines[2:]:
            parts = line.strip().split("\t")
            state = tuple(map(int, parts[0].split()))
            path_data.append(state)
    return path_data

# === DQN网络 ===
class DQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, 128)
        self.fc2 = nn.Linear(128, 64)
        self.fc3 = nn.Linear(64, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)

# === DQN Agent with PER ===
class DQNAgentWithPER:
    def __init__(self, state_dim, action_dim, replay_buffer, gamma=0.99, epsilon=1.0, epsilon_decay=0.995,
                 epsilon_min=0.1, learning_rate=0.001, alpha=0.6, beta=0.4):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer
        self.alpha = alpha
        self.beta = beta

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        delta_values = [1, 2, 3, 5, -1, -2, -3, -5]
        dim = action_idx // 8
        delta_idx = action_idx % 8
        delta = delta_values[delta_idx]
        if dim == 0:
            return (delta, 0, 0)
        elif dim == 1:
            return (0, delta, 0)
        elif dim == 2:
            return (0, 0, delta)

    def act(self, state):
        if random.random() < self.epsilon:
            return random.randrange(self.action_dim)
        normalized_state = normalize_state(state)
        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)
        return torch.argmax(q_values, dim=1).item()

    def store_transition(self, state, action, reward, next_state, done):
        normalized_state = normalize_state(state)
        normalized_next_state = normalize_state(next_state)
        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        next_state_tensor = torch.tensor(normalized_next_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)
            next_q_values = self.target_model(next_state_tensor)
            max_next_q_values = next_q_values.max(1)[0]
            target_q_values = reward + (self.gamma * max_next_q_values * (1 - done))
            td_error = torch.abs(q_values[0][action] - target_q_values).item()
        self.replay_buffer.append((state_tensor, action, reward, next_state_tensor, done, td_error))
        return td_error

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return
        batch, batch_indices, probabilities = self.replay_buffer.sample(batch_size, alpha=self.alpha)
        states, actions, rewards, next_states, dones, _ = zip(*batch)
        weights = (len(self.replay_buffer) * probabilities) ** (-self.beta)
        weights = weights / weights.max()
        weights = torch.tensor(weights, dtype=torch.float32).to(device)

        states = torch.tensor(np.array([s.cpu().numpy().flatten() for s in states]), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array([ns.cpu().numpy().flatten() for ns in next_states]), dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))

        td_errors = current_q_values - target_q_values
        weighted_loss = (td_errors.pow(2) * weights).mean()

        self.optimizer.zero_grad()
        weighted_loss.backward()
        self.optimizer.step()

        new_priorities = torch.abs(td_errors).detach().cpu().numpy()
        self.replay_buffer.update_priorities(batch_indices, new_priorities)

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())

# === 训练函数（组） ===
def train_group(group_paths, path_documents, replay_buffer, batch_size=32, group_name=""):
    state_dim = 3
    action_dim = 24  # 3 x 8 delta
    agent = DQNAgentWithPER(state_dim, action_dim, replay_buffer)

    global_steps = 0
    path_rewards = {}

    print(f"Start training {group_name}, Included Paths: {[idx + 1 for idx in group_paths]}")
    start_time = time.time()

    N_SAMPLES = 200
    BATCH_SIZE = 50
    N_BATCHES = 4
    N_STEPS = 3
    N_REPEATS = 5
    TARGET_UPDATE_EVERY_N_BATCHES = 2

    for path_idx in group_paths:
        file_path = os.path.join(path_documents,
                                 f"path{path_idx + 1}_{'similar' if group_name == '' else 'isolated'}.txt")
        if not os.path.exists(file_path):
            print(f"  : Path {path_idx + 1} file not found, skipping")
            continue

        path_data = load_path_data(file_path)
        target_path = target_paths[path_idx]

        if path_idx not in path_rewards:
            path_rewards[path_idx] = 0

        print(f"\n  Start training path {path_idx + 1}, samples: {len(path_data)}")

        for repeat in range(N_REPEATS):
            print(f"    Run {repeat + 1}/{N_REPEATS}")

            batch_count = 0
            for batch_idx in range(N_BATCHES):
                batch_start = batch_idx * BATCH_SIZE
                batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)

                if batch_start >= len(path_data):
                    print(f"      Run {batch_idx + 1}: insufficient samples, stopping")
                    break

                print(f"      Run {batch_idx + 1}/{N_BATCHES} (samples {batch_start}-{batch_end})")

                for sample_idx in range(batch_start, batch_end):
                    if sample_idx >= len(path_data):
                        break

                    state = path_data[sample_idx]
                    prev_state = None
                    prev_triggered = None

                    for step in range(N_STEPS):
                        legal_actions = []
                        for a in range(agent.action_dim):
                            dw, dt, dz = agent.decode_action(a)
                            cand_next = (state[0] + dw, state[1] + dt, state[2] + dz)
                            if (MIN_X <= cand_next[0] <= MAX_X and
                                MIN_Y <= cand_next[1] <= MAX_Y and
                                MIN_Z <= cand_next[2] <= MAX_Z):
                                legal_actions.append(a)

                        if not legal_actions:
                            break

                        if random.random() < agent.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            normalized_state = normalize_state(state)
                            state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        dw, dt, dz = agent.decode_action(action)
                        next_state = (state[0] + dw, state[1] + dt, state[2] + dz)

                        triggered = section9_hybrid_speed_torque_current(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered,
                                                prev_triggered, prev_state)

                        done = (step == N_STEPS - 1)

                        td_error = agent.store_transition(state, action, reward, next_state, done)

                        prev_state = state
                        prev_triggered = triggered
                        state = next_state

                        path_rewards[path_idx] += reward
                        global_steps += 1

                print(f"        Batch {batch_idx + 1} completed, training...")
                if len(agent.replay_buffer) >= batch_size:
                    agent.train(batch_size)

                batch_count += 1

                if batch_count % TARGET_UPDATE_EVERY_N_BATCHES == 0:
                    agent.update_target_model()
                    print(f"        Target model updated at batch {batch_count}")

        print(f"  Path {path_idx + 1} completed, total reward: {path_rewards[path_idx]:.2f}")

    training_time = time.time() - start_time
    print(f"\n{group_name} training completed, time: {training_time:.2f} seconds")
    print(f"Replay buffer size: {len(replay_buffer)}")

    return agent, path_rewards, training_time

# === 分阶段生成与训练 ===
def generate_and_train_grouped_paths_staged(path_documents, similar_group, isolated_group, batch_size=32, run_id=1):
    print(f"\n=== Run {run_id}/20 ===")
    similar_group_paths = [idx + 1 for idx in similar_group]
    isolated_group_paths = [idx + 1 for idx in isolated_group]

    print(f"Similar group paths: {similar_group_paths}")
    print(f"Isolated group paths: {isolated_group_paths}")

    total_start_time = time.time()

    print(f"\n[1] Generating samples for similar group...")
    generate_samples_for_similar_paths(similar_group, num_candidates=2000, top_k=200, run_id=run_id)

    print(f"\n[2] Training on similar group...")
    similar_replay_buffer = GroupExperienceReplay(capacity=20000)
    similar_agent, similar_path_rewards, similar_training_time = train_group(
        similar_group, path_documents, similar_replay_buffer, batch_size=batch_size, group_name="similar"
    )

    print(f"\n[3] Generating samples for isolated group using similar model...")
    generate_samples_for_isolated_paths(isolated_group, similar_agent.model, num_candidates=2000, top_k=200,
                                        run_id=run_id)

    print(f"\n[4] Training on isolated group...")
    isolated_replay_buffer = GroupExperienceReplay(capacity=20000)
    isolated_agent, isolated_path_rewards, isolated_training_time = train_group(
        isolated_group, path_documents, isolated_replay_buffer, batch_size=batch_size, group_name="isolated"
    )

    total_path_rewards = {**similar_path_rewards, **isolated_path_rewards}
    total_cumulative_reward = sum(total_path_rewards.values())
    total_training_time = time.time() - total_start_time

    print(f"\n=== Run {run_id}/20 completed, total time: {total_training_time:.2f} seconds ===")

    return similar_agent, isolated_agent, similar_replay_buffer, isolated_replay_buffer, total_cumulative_reward, total_path_rewards, total_training_time

# === Excel 报告生成 ===
def create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir):
    os.makedirs(output_dir, exist_ok=True)

    similar_group_paths = [idx + 1 for idx in similar_group]
    isolated_group_paths = [idx + 1 for idx in isolated_group]

    wb = Workbook()

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    header_color = "4472C4"
    similar_group_color = "E2EFDA"
    isolated_group_color = "FCE4D6"
    stats_color = "FFF2CC"

    # === 1: 路径相似度汇总 ===
    ws_paths = wb.active
    ws_paths.title = "Path Similarities"

    path_headers = ['Path ID', 'Group'] + [f'Run {i}' for i in range(1, 21)] + ['Average', 'Max', 'Min', 'Std']
    for col, header in enumerate(path_headers, 1):
        cell = ws_paths.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_paths.row_dimensions[1].height = 30

    for path_id in range(1, len(target_paths) + 1):
        row = path_id + 1
        if path_id in similar_group_paths:
            group_type = "High-correlation"
            row_color = similar_group_color
        elif path_id in isolated_group_paths:
            group_type = "Low-correlation"
            row_color = isolated_group_color
        else:
            group_type = "Ungrouped"
            row_color = "FFFFFF"

        ws_paths.cell(row=row, column=1, value=f"Path {path_id}").font = Font(bold=True)
        ws_paths.cell(row=row, column=2, value=group_type)
        for col, cell in enumerate([ws_paths.cell(row=row, column=1), ws_paths.cell(row=row, column=2)], start=1):
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
            cell.border = thin_border

        path_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            sim = run_data['path_similarities'].get(path_id, {}).get('avg_similarity', 0.0)
            path_similarities.append(sim)
            cell = ws_paths.cell(row=row, column=3 + run_idx, value=round(sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        stats_values = [round(np.mean(path_similarities), 4),
                        round(np.max(path_similarities), 4),
                        round(np.min(path_similarities), 4),
                        round(np.std(path_similarities), 4)]
        for i, val in enumerate(stats_values):
            cell = ws_paths.cell(row=row, column=23 + i, value=val)
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
            cell.font = Font(bold=True)
            cell.border = thin_border

    for col in range(1, 27):
        ws_paths.column_dimensions[get_column_letter(col)].width = 12

    # === 2: 分组统计 ===
    ws_groups = wb.create_sheet("Group Statistics")
    group_headers = ['Group Name', 'Included Paths'] + [f'Run {i}' for i in range(1, 21)] + ['Average', 'Std']
    for col, header in enumerate(group_headers, 1):
        cell = ws_groups.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws_groups.row_dimensions[1].height = 30

    row = 2
    # Similar group
    cell = ws_groups.cell(row=row, column=1, value="High-correlation group")
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    ws_groups.cell(row=row, column=2, value=','.join(map(str, similar_group_paths))).fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    for col, cell in enumerate([ws_groups.cell(row=row, column=1), ws_groups.cell(row=row, column=2)], start=1):
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    group_sims = []
    for run_idx, run_data in enumerate(all_runs_data):
        sim = np.mean([run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0) for p in similar_group_paths])
        group_sims.append(sim)
        cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(sim, 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws_groups.cell(row=row, column=23, value=round(np.mean(group_sims), 4)).number_format = '0.0000'
    ws_groups.cell(row=row, column=24, value=round(np.std(group_sims), 4)).number_format = '0.0000'
    for col in [23, 24]:
        cell = ws_groups.cell(row=row, column=col)
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = thin_border

    row += 1
    # Isolated group
    if isolated_group_paths:
        cell = ws_groups.cell(row=row, column=1, value="Low-correlation group")
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        ws_groups.cell(row=row, column=2, value=','.join(map(str, isolated_group_paths))).fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        for col, cell in enumerate([ws_groups.cell(row=row, column=1), ws_groups.cell(row=row, column=2)], start=1):
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        iso_sims = []
        for run_idx, run_data in enumerate(all_runs_data):
            sim = np.mean([run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0) for p in isolated_group_paths])
            iso_sims.append(sim)
            cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws_groups.cell(row=row, column=23, value=round(np.mean(iso_sims), 4)).number_format = '0.0000'
        ws_groups.cell(row=row, column=24, value=round(np.std(iso_sims), 4)).number_format = '0.0000'
        for col in [23, 24]:
            cell = ws_groups.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
            cell.font = Font(bold=True)
            cell.border = thin_border

    for col in range(1, 25):
        ws_groups.column_dimensions[get_column_letter(col)].width = 12

    # === 3: 详细样本数据 ===
    ws_samples = wb.create_sheet("Detailed Samples")
    sample_headers = ['Run', 'Path ID', 'Sample ID', 'Weather', 'TimePeriod', 'Z', 'Similarity', 'Triggered Rules']
    for col, header in enumerate(sample_headers, 1):
        cell = ws_samples.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws_samples.row_dimensions[1].height = 30

    sample_row = 2
    for run_idx, run_data in enumerate(all_runs_data, 1):
        for path_id in range(1, len(target_paths) + 1):
            samples = run_data['path_samples'].get(path_id, [])
            if not samples:
                continue
            for sample_idx, (state_tuple, reward, sim, triggered) in enumerate(samples, 1):
                weather, time_period, z = state_tuple
                triggered_str = ','.join(map(str, sorted(triggered)))
                ws_samples.cell(row=sample_row, column=1, value=f"Run {run_idx}")
                ws_samples.cell(row=sample_row, column=2, value=f"Path {path_id}")
                ws_samples.cell(row=sample_row, column=3, value=sample_idx)
                ws_samples.cell(row=sample_row, column=4, value=weather)
                ws_samples.cell(row=sample_row, column=5, value=time_period)
                ws_samples.cell(row=sample_row, column=6, value=z)
                ws_samples.cell(row=sample_row, column=7, value=round(sim, 4)).number_format = '0.0000'
                ws_samples.cell(row=sample_row, column=8, value=f"{{{triggered_str}}}")
                for col in range(1, 9):
                    cell = ws_samples.cell(row=sample_row, column=col)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                sample_row += 1

    sample_widths = [13, 13, 11, 10, 12, 8, 12, 50]
    for i, width in enumerate(sample_widths, 1):
        ws_samples.column_dimensions[get_column_letter(i)].width = width

    output_path = os.path.join(output_dir, "20_runs_consolidated_report.xlsx")
    wb.save(output_path)
    print(f"\nConsolidated Excel report saved to: {output_path}")

# === 主运行函数（20次） ===
def run_20_times_training():
    """执行20次完整训练流程"""
    model_path_base = r"D:\Experiment\CNN\DQNNEW\saved_models_traffic"
    path_documents = r"D:\Experiment\CNN\DQNNEW\path_samples_grouped"
    output_dir = r"D:\Experiment\CNN\ComparisonExperiment2\excel_reports_traffic"

    os.makedirs(model_path_base, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    similar_group, isolated_group = group_paths_by_similarity(target_paths)
    similar_group_display = [idx + 1 for idx in similar_group]
    isolated_group_display = [idx + 1 for idx in isolated_group]

    print("=" * 60)
    print("20-run Training Experiment")
    print(f"Value ranges: weather[{MIN_X},{MAX_X}], time_period[{MIN_Y},{MAX_Y}], z[{MIN_Z},{MAX_Z}]")
    print("Process: Generate samples → Train similar group → Generate isolated samples → Train isolated group")
    print("=" * 60)
    print(f"\nAutomatic grouping results:")
    print(f"Similar group (high-correlation): {similar_group_display}")
    print(f"Isolated group (low-correlation): {isolated_group_display}")
    print("\n" + "=" * 60)

    all_runs_data = []
    total_start_time = time.time()

    for run_id in range(1, 21):
        print(f"\n{'=' * 60}")
        print(f"Starting Run {run_id}/20")
        print(f"{'=' * 60}")

        similar_agent, isolated_agent, similar_buffer, isolated_buffer, total_cumulative_reward, path_rewards, training_time = \
            generate_and_train_grouped_paths_staged(path_documents, similar_group, isolated_group, batch_size=32,
                                                    run_id=run_id)

        # 保存模型
        similar_model_path = os.path.join(model_path_base, f"similar_group_model_run_{run_id}.pth")
        isolated_model_path = os.path.join(model_path_base, f"isolated_group_model_run_{run_id}.pth")

        torch.save({
            'model_state_dict': similar_agent.model.state_dict(),
            'optimizer_state_dict': similar_agent.optimizer.state_dict(),
            'epsilon': similar_agent.epsilon,
            'normalization': {'x_range': (MIN_X, MAX_X), 'y_range': (MIN_Y, MAX_Y), 'z_range': (MIN_Z, MAX_Z)},
            'run_id': run_id,
            'group_type': 'similar_group',
            'group_paths': similar_group_display,
            'pool_size': len(similar_buffer),
            'pool_capacity': 20000,
        }, similar_model_path)

        torch.save({
            'model_state_dict': isolated_agent.model.state_dict(),
            'optimizer_state_dict': isolated_agent.optimizer.state_dict(),
            'epsilon': isolated_agent.epsilon,
            'normalization': {'x_range': (MIN_X, MAX_X), 'y_range': (MIN_Y, MAX_Y), 'z_range': (MIN_Z, MAX_Z)},
            'run_id': run_id,
            'group_type': 'isolated_group',
            'group_paths': isolated_group_display,
            'pool_size': len(isolated_buffer),
            'pool_capacity': 20000,
        }, isolated_model_path)

        print(f"[Run {run_id}] Models saved.")

        run_data = {
            'run_id': run_id,
            'training_time': training_time,
            'total_reward': total_cumulative_reward,
            'path_rewards': path_rewards,
            'path_similarities': {},
            'path_samples': {}
        }

        all_similarities = []
        for path_idx in range(len(target_paths)):
            path_id = path_idx + 1
            target_path = target_paths[path_idx]

            if path_id in similar_group_display:
                buffer = similar_buffer
            elif path_id in isolated_group_display:
                buffer = isolated_buffer
            else:
                continue

            high_reward_samples = buffer.get_high_reward_samples(target_path, num_samples=20)
            if high_reward_samples:
                similarities = [sim for _, _, sim, _ in high_reward_samples]
                run_data['path_similarities'][path_id] = {
                    'avg_similarity': np.mean(similarities),
                    'max_similarity': np.max(similarities),
                    'min_similarity': np.min(similarities),
                    'sample_count': len(similarities)
                }
                run_data['path_samples'][path_id] = high_reward_samples
                all_similarities.extend(similarities)
            else:
                run_data['path_similarities'][path_id] = {
                    'avg_similarity': 0.0,
                    'max_similarity': 0.0,
                    'min_similarity': 0.0,
                    'sample_count': 0
                }
                run_data['path_samples'][path_id] = []

        run_data['overall_avg_similarity'] = np.mean(all_similarities) if all_similarities else 0.0
        all_runs_data.append(run_data)

        print(f"[Run {run_id}] Completed! Overall Avg Similarity: {run_data['overall_avg_similarity']:.4f}")
        print(f"{'=' * 60}\n")

    total_time = time.time() - total_start_time

    print("\nGenerating consolidated Excel report...")
    create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir)

    print("\n" + "=" * 60)
    print("All 20 runs completed! Summary:")
    print("=" * 60)
    print(f"Value ranges:")
    print(f"  weather (X): [{MIN_X}, {MAX_X}]")
    print(f"  time_period (Y): [{MIN_Y}, {MAX_Y}]")
    print(f"  z (Z): [{MIN_Z}, {MAX_Z}]")
    print(f"\nTotal elapsed time: {total_time:.2f} seconds ({total_time/60:.2f} minutes)")
    print(f"Average time per run: {total_time/20:.2f} seconds")
    print(f"\nAverage similarity statistics over 20 runs:")
    avg_sims = [r['overall_avg_similarity'] for r in all_runs_data]
    print(f"  Overall mean: {np.mean(avg_sims):.4f}")
    print(f"  Maximum: {np.max(avg_sims):.4f}")
    print(f"  Minimum: {np.min(avg_sims):.4f}")
    print(f"  Std deviation: {np.std(avg_sims):.4f}")
    print(f"\nAll results saved to: {output_dir}")
    print("=" * 60)

# 执行主程序
if __name__ == "__main__":
    run_20_times_training()