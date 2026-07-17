import os
import random
import math
import numpy as np
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# ================================================================
# 1. 状态范围配置（按您的最新要求）
# ================================================================
STATE_MIN_X, STATE_MAX_X = 140, 240   # 温度/速度
STATE_MIN_Y, STATE_MAX_Y = 200, 250   # 电压/扭矩
STATE_MIN_Z, STATE_MAX_Z = 1, 90      # 流量/电流

def generate_input():
    # 使用 randint 生成整数，因为 section9 规则大多为整数比较
    return [
        random.randint(STATE_MIN_X, STATE_MAX_X),
        random.randint(STATE_MIN_Y, STATE_MAX_Y),
        random.randint(STATE_MIN_Z, STATE_MAX_Z)
    ]

# ================================================================
# 2. 规则函数 section9（您提供的完整函数）
# ================================================================
def section9_hybrid_speed_torque_current(x, y, z):
    triggered = set()
    speed_std = (1000, 2000)
    torque_std = (100, 200)
    current_std = (10, 30)

    if [(160 < x < 240)] != [(160 < x * 8 < 240)]:
        triggered.add(1)
    if [(160 < x < 240)] != [(160 < 100 < 240)]:
        triggered.add(2)

    if [(210 < y < 250)] != [(210 < y * 9 < 250)]:
        triggered.add(3)
    if [(210 < y < 250)] != [(210 < y * 12 < 250)]:
        triggered.add(4)

    if [(40 < z < 90)] != [(40 < z * 8 < 90)]:
        triggered.add(5)
    if [(40 < z < 90)] != [(40 < z * 9 < 90)]:
        triggered.add(6)

    if [(x > 190 and x < 210 and y > 225 and y < 235)] != [(x > 190 and x < 210 and y > 225 and 200 < 235)]:
        triggered.add(7)
    if [(x > 190 and x < 210 and y > 225 and y < 235)] != [(x > 190 and x < 210 and y > 225 and 220 < 235)]:
        triggered.add(8)

    if [(x > 185 and x < 215 and z > 60 and z < 70)] != [(x > 185 and x < 215 and z > 60 and 60 < 70)]:
        triggered.add(9)
    if [(x > 185 and x < 215 and z > 60 and z < 70)] != [(x > 185 and x < 215 and z > 60 and 55 < 70)]:
        triggered.add(10)

    if [(y > 223 and y < 237 and z > 60 and z < 70)] != [(y > 223 and y < 237 and z > 60 and 60 < 70)]:
        triggered.add(11)
    if [(y > 223 and y < 237 and z > 60 and z < 70)] != [(y > 223 and y < 237 and z > 60 and 67 < 70)]:
        triggered.add(12)

    if [(x > 198 and x < 202)] != [(x > 198 and x * 8 < 202)]:
        triggered.add(13)
    if [(x > 198 and x < 202)] != [(x > 198 and 200 < 202)]:
        triggered.add(14)

    if [(y > 228 and y < 232)] != [(y > 228 and y * 8 < 232)]:
        triggered.add(15)
    if [(y > 228 and y < 232)] != [(y > 228 and 200 < 232)]:
        triggered.add(16)

    if [(z > 63 and z < 67)] != [(z > 63 and 44 < 67)]:
        triggered.add(17)
    if [(z > 63 and z < 67)] != [(z > 63 and 56 < 67)]:
        triggered.add(18)

    if [(abs(x - 200) < 5)] != [(abs(x - 200) < 51)]:
        triggered.add(19)
    if [(abs(x - 200) < 5)] != [(abs(x - 200) < 15)]:
        triggered.add(20)

    if [(abs(y - 230) < 3)] != [(abs(y - 230) < 13)]:
        triggered.add(21)
    if [(abs(y - 230) < 3)] != [(abs(y - 230) < 31)]:
        triggered.add(22)

    if [(abs(z - 65) < 2)] != [(abs(z - 65) < 21)]:
        triggered.add(23)
    if [(abs(z - 65) < 2)] != [(abs(z - 65) < 12)]:
        triggered.add(24)

    if [(x > 185 and y > 223 and z > 55)] != [(x > 185 and y > 223 and z > 155)]:
        triggered.add(25)
    if [(x > 185 and y > 223 and z > 55)] != [(x > 185 and y > 223 and z > 515)]:
        triggered.add(26)

    if [(x < 215 and y < 237 and z < 75)] != [(x < 215 and y < 237 and z < 715)]:
        triggered.add(27)
    if [(x < 215 and y < 237 and z < 75)] != [(x < 215 and y < 237 and z < 225)]:
        triggered.add(28)

    if [((x + y / 10 + z) > 288 and (x + y / 10 + z) < 308)] != [((x + y / 10 + z) > 288 and (x + y / 10 + z) < 508)]:
        triggered.add(29)
    if [((x + y / 10 + z) > 288 and (x + y / 10 + z) < 308)] != [((x + y / 10 + z) > 288 and (x + y / 10 + z) < 708)]:
        triggered.add(30)

    if [(x * y / 100 > 440 and x * y / 100 < 500)] != [(x * y / 100 > 440 and x * y / 100 < 900)]:
        triggered.add(31)
    if [(x * y / 100 > 440 and x * y / 100 < 500)] != [(x * y / 100 > 440 and x * y / 100 < 440)]:
        triggered.add(32)

    if [(x * z > 12500 and x * z < 14500)] != [(x * z > 12500 and x * z < 17600)]:
        triggered.add(33)
    if [(x * z > 12500 and x * z < 14500)] != [(x * z > 12500 and x * z < 54500)]:
        triggered.add(34)

    if [(y / 10 * z > 1400 and y / 10 * z < 1600)] != [(y / 10 * z > 1400 and y / 10 * z < 7800)]:
        triggered.add(35)
    if [(y / 10 * z > 1400 and y / 10 * z < 1600)] != [(y / 10 * z > 1400 and y / 10 * z < 5600)]:
        triggered.add(36)

    if [((x + y / 10 + z) / 3 > 96 and (x + y / 10 + z) / 3 < 102)] != [((x + y / 10 + z) / 3 > 96 and (x + y / 10 + z) / 3 < 1052)]:
        triggered.add(37)
    if [((x + y / 10 + z) / 3 > 96 and (x + y / 10 + z) / 3 < 102)] != [((x + y / 10 + z) / 3 > 96 and (x + y / 10 + z) / 3 < 1502)]:
        triggered.add(38)

    if [(math.sqrt(x ** 2 + (y / 10) ** 2 + z ** 2) > 220)] != [(math.sqrt(x ** 2 + (y / 10) ** 2 + z ** 2) > 2520)]:
        triggered.add(39)
    if [(math.sqrt(x ** 2 + (y / 10) ** 2 + z ** 2) > 220)] != [(math.sqrt(x ** 2 + (y / 10) ** 2 + z ** 2) > 280)]:
        triggered.add(40)

    if [(x / (y / 10) > 8.2 and x / (y / 10) < 9.2)] != [(x / (y / 10) > 8.2 and x / (y / 10) < 92)]:
        triggered.add(41)
    if [(x / (y / 10) > 8.2 and x / (y / 10) < 9.2)] != [(x / (y / 10) > 8.2 and x / (y / 10) < 19.2)]:
        triggered.add(42)

    if [(x / z > 2.8 and x / z < 3.4)] != [(x / z > 2.8 and x / z < 13.4)]:
        triggered.add(43)
    if [(x / z > 2.8 and x / z < 3.4)] != [(x / z > 2.8 and x / z < 31.4)]:
        triggered.add(44)

    if [((y / 10) / z > 0.32 and (y / 10) / z < 0.38)] != [((y / 10) / z > 0.32 and (y / 10) / z < 38)]:
        triggered.add(45)
    if [((y / 10) / z > 0.32 and (y / 10) / z < 0.38)] != [((y / 10) / z > 0.32 and (y / 10) / z < 3.8)]:
        triggered.add(46)

    if [((x - 180) + (y - 220) / 10 > 22 and (x - 180) + (y - 220) / 10 < 32)] != [((x - 180) + (y - 220) / 10 > 22 and (x - 180) + (y - 220) / 10 < 312)]:
        triggered.add(47)
    if [((x - 180) + (y - 220) / 10 > 22 and (x - 180) + (y - 220) / 10 < 32)] != [((x - 180) + (y - 220) / 10 > 22 and (x - 180) + (y - 220) / 10 < 132)]:
        triggered.add(48)

    if [((x - 180) + (z - 50) > 35 and (x - 180) + (z - 50) < 45)] != [((x - 180) + (z - 50) > 35 and (x - 180) + (z - 50) < 415)]:
        triggered.add(49)
    if [((x - 180) + (z - 50) > 35 and (x - 180) + (z - 50) < 45)] != [((x - 180) + (z - 50) > 35 and (x - 180) + (z - 50) < 145)]:
        triggered.add(50)

    if [((y - 220) / 10 + (z - 50) > 25 and (y - 220) / 10 + (z - 50) < 35)] != [((y - 220) / 10 + (z - 50) > 25 and (y - 220) / 10 + (z - 50) < 315)]:
        triggered.add(51)
    if [((y - 220) / 10 + (z - 50) > 25 and (y - 220) / 10 + (z - 50) < 35)] != [((y - 220) / 10 + (z - 50) > 25 and (y - 220) / 10 + (z - 50) < 135)]:
        triggered.add(52)

    if [(abs((x - 200) - (y - 230) / 2) < 8)] != [(abs((x - 200) - (y - 230) / 2) < 18)]:
        triggered.add(53)
    if [(abs((x - 200) - (y - 230) / 2) < 8)] != [(abs((x - 200) - (y - 230) / 2) < 81)]:
        triggered.add(54)

    if [(abs((x - 200) - (z - 65) * 3) < 12)] != [(abs((x - 200) - (z - 65) * 3) < 92)]:
        triggered.add(55)
    if [(abs((x - 200) - (z - 65) * 3) < 12)] != [(abs((x - 200) - (z - 65) * 3) < 122)]:
        triggered.add(56)

    if [(abs((y - 230) / 10 - (z - 65) / 5) < 3)] != [(abs((y - 230) / 10 - (z - 65) / 5) < 13)]:
        triggered.add(57)
    if [(abs((y - 230) / 10 - (z - 65) / 5) < 3)] != [(abs((y - 230) / 10 - (z - 65) / 5) < 31)]:
        triggered.add(58)

    if [(x / (y / 10 + 5) > 6.5 and x / (y / 10 + 5) < 8.5)] != [(x / (y / 10 + 5) > 6.5 and x / (y / 10 + 5) < 85)]:
        triggered.add(59)
    if [(x / (y / 10 + 5) > 6.5 and x / (y / 10 + 5) < 8.5)] != [(x / (y / 10 + 5) > 6.5 and x / (y / 10 + 5) < 81.5)]:
        triggered.add(60)

    if [((y / 10) / (z + 15) > 0.26 and (y / 10) / (z + 15) < 0.34)] != [((y / 10) / (z + 15) > 0.26 and (y / 10) / (z + 15) < 10.34)]:
        triggered.add(61)
    if [((y / 10) / (z + 15) > 0.26 and (y / 10) / (z + 15) < 0.34)] != [((y / 10) / (z + 15) > 0.26 and (y / 10) / (z + 15) < 40.34)]:
        triggered.add(62)

    if [(z / (x / 5) > 1.4 and z / (x / 5) < 1.8)] != [(z / (x / 5) > 1.4 and z / (x / 5) < 11.8)]:
        triggered.add(63)
    if [(z / (x / 5) > 1.4 and z / (x / 5) < 1.8)] != [(z / (x / 5) > 1.4 and z / (x / 5) < 41.8)]:
        triggered.add(64)

    if [((x * y * z / 10000) > 28 and (x * y * z / 10000) < 36)] != [((x * y * z / 10000) > 28 and (x * y * z / 10000) < 332)]:
        triggered.add(65)
    if [((x * y * z / 10000) > 28 and (x * y * z / 10000) < 36)] != [((x * y * z / 10000) > 28 and (x * y * z / 10000) < 236)]:
        triggered.add(66)

    if [(x * 0.5 + y / 10 * 0.3 + z * 0.2 > 125 and x * 0.5 + y / 10 * 0.3 + z * 0.2 < 135)] != [(x * 0.5 + y / 10 * 0.3 + z * 0.2 > 125 and x * 0.5 + y / 10 * 0.3 + z * 0.2 < 935)]:
        triggered.add(67)
    if [(x * 0.5 + y / 10 * 0.3 + z * 0.2 > 125 and x * 0.5 + y / 10 * 0.3 + z * 0.2 < 135)] != [(x * 0.5 + y / 10 * 0.3 + z * 0.2 > 125 and x * 0.5 + y / 10 * 0.3 + z * 0.2 < 635)]:
        triggered.add(68)

    if [((x / 200) ** 0.4 * (y / 230) ** 0.3 * (z / 65) ** 0.3 > 0.92)] != [((x / 200) ** 0.4 * (y / 230) ** 0.3 * (z / 65) ** 0.3 > 9.2)]:
        triggered.add(69)
    if [((x / 200) ** 0.4 * (y / 230) ** 0.3 * (z / 65) ** 0.3 > 0.92)] != [((x / 200) ** 0.4 * (y / 230) ** 0.3 * (z / 65) ** 0.3 > 92)]:
        triggered.add(70)

    if [((x - 200) * (y - 230) / 10 > -80 and (x - 200) * (y - 230) / 10 < 80)] != [((x - 200) * (y - 230) / 10 > -80 and (x - 200) * (y - 230) / 10 < 180)]:
        triggered.add(71)
    if [((x - 200) * (y - 230) / 10 > -80 and (x - 200) * (y - 230) / 10 < 80)] != [((x - 200) * (y - 230) / 10 > -80 and (x - 200) * (y - 230) / 10 < 810)]:
        triggered.add(72)

    if [((x - 200) * (z - 65) > -120 and (x - 200) * (z - 65) < 120)] != [((x - 200) * (z - 65) > -120 and (x - 200) * (z - 65) < 820)]:
        triggered.add(73)
    if [((x - 200) * (z - 65) > -120 and (x - 200) * (z - 65) < 120)] != [((x - 200) * (z - 65) > -120 and (x - 200) * (z - 65) < 520)]:
        triggered.add(74)

    if [((y - 230) / 10 * (z - 65) > -30 and (y - 230) / 10 * (z - 65) < 30)] != [((y - 230) / 10 * (z - 65) > -30 and (y - 230) / 10 * (z - 65) < 310)]:
        triggered.add(75)
    if [((y - 230) / 10 * (z - 65) > -30 and (y - 230) / 10 * (z - 65) < 30)] != [((y - 230) / 10 * (z - 65) > -30 and (y - 230) / 10 * (z - 65) < 130)]:
        triggered.add(76)

    if [(abs(x / 200 + y / 230 + z / 65 - 3) < 0.18)] != [(abs(x / 200 + y / 230 + z / 65 - 3) < 18)]:
        triggered.add(77)
    if [(abs(x / 200 + y / 230 + z / 65 - 3) < 0.18)] != [(abs(x / 200 + y / 230 + z / 65 - 3) < 10.18)]:
        triggered.add(78)

    if [(abs((x / 200) * (y / 230) * (z / 65) - 1) < 0.12)] != [(abs((x / 200) * (y / 230) * (z / 65) - 1) < 10.12)]:
        triggered.add(79)
    if [(abs((x / 200) * (y / 230) * (z / 65) - 1) < 0.12)] != [(abs((x / 200) * (y / 230) * (z / 65) - 1) < 12.12)]:
        triggered.add(80)

    if [(x > 197 and x < 203 and y > 228 and y < 232 and z > 63.5 and z < 66.5)] != [(x > 197 and x < 203 and y > 228 and y < 232 and z > 63.5 and z < 616.5)]:
        triggered.add(81)
    if [(x > 197 and x < 203 and y > 228 and y < 232 and z > 63.5 and z < 66.5)] != [(x > 197 and x < 203 and y > 228 and y < 232 and z > 63.5 and z < 166.5)]:
        triggered.add(82)

    if [(abs(x / (y / 10) - 8.7) < 0.3 and abs(z - 65) < 1.5)] != [(abs(x / (y / 10) - 8.7) < 0.3 and abs(z - 65) < 11.5)]:
        triggered.add(83)
    if [(abs(x / (y / 10) - 8.7) < 0.3 and abs(z - 65) < 1.5)] != [(abs(x / (y / 10) - 8.7) < 0.3 and abs(z - 65) < 15)]:
        triggered.add(84)

    if [(x * y / 100 > 455 and x * y / 100 < 465 and z > 64 and z < 66)] != [(x * y / 100 > 455 and x * y / 100 < 465 and z > 64 and z < 166)]:
        triggered.add(85)
    if [(x * y / 100 > 455 and x * y / 100 < 465 and z > 64 and z < 66)] != [(x * y / 100 > 455 and x * y / 100 < 465 and z > 64 and z < 616)]:
        triggered.add(86)

    if [((x + y / 10) / 2 > 112 and (x + y / 10) / 2 < 116 and z > 63.5 and z < 66.5)] != [((x + y / 10) / 2 > 112 and (x + y / 10) / 2 < 116 and z > 63.5 and z < 166.5)]:
        triggered.add(87)
    if [((x + y / 10) / 2 > 112 and (x + y / 10) / 2 < 116 and z > 63.5 and z < 66.5)] != [((x + y / 10) / 2 > 112 and (x + y / 10) / 2 < 116 and z > 63.5 and z < 661.5)]:
        triggered.add(88)

    if [(abs(x - y / 10) < 175 and z > 63.5 and z < 66.5)] != [(abs(x - y / 10) < 175 and z > 63.5 and z < 616.5)]:
        triggered.add(89)
    if [(abs(x - y / 10) < 175 and z > 63.5 and z < 66.5)] != [(abs(x - y / 10) < 175 and z > 63.5 and z < 6615)]:
        triggered.add(90)

    if [(math.sqrt((x - 200) ** 2 + (y / 10 - 23) ** 2) < 3 and abs(z - 65) < 1.5)] != [(math.sqrt((x - 200) ** 2 + (y / 10 - 23) ** 2) < 3 and abs(z - 65) < 15)]:
        triggered.add(91)
    if [(math.sqrt((x - 200) ** 2 + (y / 10 - 23) ** 2) < 3 and abs(z - 65) < 1.5)] != [(math.sqrt((x - 200) ** 2 + (y / 10 - 23) ** 2) < 3 and abs(z - 65) < 11.5)]:
        triggered.add(92)

    if [(x / 200 > 0.985 and x / 200 < 1.015 and y / 230 > 0.987 and y / 230 < 1.013)] != [(x / 200 > 0.985 and x / 200 < 1.015 and y / 230 > 0.987 and y / 230 < 10.13)]:
        triggered.add(93)
    if [(x / 200 > 0.985 and x / 200 < 1.015 and y / 230 > 0.987 and y / 230 < 1.013)] != [(x / 200 > 0.985 and x / 200 < 1.015 and y / 230 > 0.987 and y / 230 < 1013)]:
        triggered.add(94)

    if [(z / 65 > 0.985 and z / 65 < 1.015)] != [(z / 65 > 0.985 and z / 65 < 121.015)]:
        triggered.add(95)
    if [(z / 65 > 0.985 and z / 65 < 1.015)] != [(z / 65 > 0.985 and z / 65 < 111.015)]:
        triggered.add(96)

    if [((x / 200 + y / 230 + z / 65) / 3 > 0.985 and (x / 200 + y / 230 + z / 65) / 3 < 1.015)] != [((x / 200 + y / 230 + z / 65) / 3 > 0.985 and (x / 200 + y / 230 + z / 65) / 3 < 2.115)]:
        triggered.add(97)
    if [((x / 200 + y / 230 + z / 65) / 3 > 0.985 and (x / 200 + y / 230 + z / 65) / 3 < 1.015)] != [((x / 200 + y / 230 + z / 65) / 3 > 0.985 and (x / 200 + y / 230 + z / 65) / 3 < 15)]:
        triggered.add(98)

    if [(max(abs(x / 200 - 1), abs(y / 230 - 1), abs(z / 65 - 1)) < 0.02)] != [(max(abs(x / 200 - 1), abs(y / 230 - 1), abs(z / 65 - 1)) < 22)]:
        triggered.add(99)
    if [(max(abs(x / 200 - 1), abs(y / 230 - 1), abs(z / 65 - 1)) < 0.02)] != [(max(abs(x / 200 - 1), abs(y / 230 - 1), abs(z / 65 - 1)) < 10.02)]:
        triggered.add(100)

    if [(min(x / 200, y / 230, z / 65) > 0.985)] != [(min(x / 200, y / 230, z / 65) > 9.85)]:
        triggered.add(101)
    if [(min(x / 200, y / 230, z / 65) > 0.985)] != [(min(x / 200, y / 230, z / 65) > 10.985)]:
        triggered.add(102)

    if [(max(x / 200, y / 230, z / 65) < 1.015)] != [(max(x / 200, y / 230, z / 65) < 10.15)]:
        triggered.add(103)
    if [(max(x / 200, y / 230, z / 65) < 1.015)] != [(max(x / 200, y / 230, z / 65) < 15)]:
        triggered.add(104)

    if [(abs(max(x, y / 10 * 8.7, z) - min(x, y / 10 * 8.7, z)) < 15)] != [(abs(max(x, y / 10 * 8.7, z) - min(x, y / 10 * 8.7, z)) < 125)]:
        triggered.add(105)
    if [(abs(max(x, y / 10 * 8.7, z) - min(x, y / 10 * 8.7, z)) < 15)] != [(abs(max(x, y / 10 * 8.7, z) - min(x, y / 10 * 8.7, z)) < 88)]:
        triggered.add(106)

    if [((x + y / 10 + z) > 297 and (x + y / 10 + z) < 299)] != [((x + y / 10 + z) > 297 and (x + y / 10 + z) < 99)]:
        triggered.add(107)
    if [((x + y / 10 + z) > 297 and (x + y / 10 + z) < 299)] != [((x + y / 10 + z) > 297 and (x + y / 10 + z) < 549)]:
        triggered.add(108)

    if [((x * y * z / 10000) > 31.5 and (x * y * z / 10000) < 345)] != [((x * y * z / 10000) > 31.5 and (x * y * z / 10000) < 132.5)]:
        triggered.add(109)
    if [((x * y * z / 10000) > 31.5 and (x * y * z / 10000) < 32.5)] != [((x * y * z / 10000) > 31.5 and (x * y * z / 10000) < 232.5)]:
        triggered.add(110)

    if [(abs((x + y / 10 + z) / 3 - 99.3) < 0.8)] != [(abs((x + y / 10 + z) / 3 - 99.3) < 8)]:
        triggered.add(111)
    if [(abs((x + y / 10 + z) / 3 - 99.3) < 0.8)] != [(abs((x + y / 10 + z) / 3 - 99.3) < 10.8)]:
        triggered.add(112)

    if [(x > 199 and x < 201 and y > 229 and y < 231)] != [(x > 199 and x < 201 and y > 229 and y < 2231)]:
        triggered.add(113)
    if [(x > 199 and x < 201 and y > 229 and y < 231)] != [(x > 199 and x < 201 and y > 229 and y < 1231)]:
        triggered.add(114)

    if [(z > 64.5 and z < 65.5)] != [(z > 64.5 and z < 615.5)]:
        triggered.add(115)
    if [(z > 64.5 and z < 65.5)] != [(z > 64.5 and z < 655)]:
        triggered.add(116)

    if [((x - 200) ** 2 + (y / 10 - 23) ** 2 + (z - 65) ** 2) < 9] != [((x - 200) ** 2 + (y / 10 - 23) ** 2 + (z - 65) ** 2) < 19]:
        triggered.add(117)
    if [((x - 200) ** 2 + (y / 10 - 23) ** 2 + (z - 65) ** 2) < 9] != [((x - 200) ** 2 + (y / 10 - 23) ** 2 + (z - 65) ** 2) < 91]:
        triggered.add(118)

    if [(abs(x / (y / 10) / z - 0.134) < 0.005)] != [(abs(x / (y / 10) / z - 0.134) < 10.005)]:
        triggered.add(119)
    if [(abs(x / (y / 10) / z - 0.134) < 0.005)] != [(abs(x / (y / 10) / z - 0.134) < 5)]:
        triggered.add(120)

    if [(x < 188 or y < 225 or z < 58)] != [(x < 188 or y < 225 or z < 158)]:
        triggered.add(121)
    if [(x < 188 or y < 225 or z < 58)] != [(x < 188 or y < 225 or 28 < 58)]:
        triggered.add(122)

    if [(x > 212 or y > 235 or z > 72)] != [(x > 212 or y > 235 or 12 * z > 72)]:
        triggered.add(123)
    if [(x > 212 or y > 235 or z > 72)] != [(x > 212 or y > 235 or 88 > 72)]:
        triggered.add(124)

    if [(abs(x / (y / 10) - 8.7) > 0.5)] != [(abs(x / (y / 10) - 8.7) > 10.5)]:
        triggered.add(125)
    if [(abs(x / (y / 10) - 8.7) > 0.5)] != [(abs(x / (y / 10) - 8.7) > 5)]:
        triggered.add(126)

    if [(abs(x / z - 3.08) > 0.3)] != [(abs(x / z - 3.08) > 10.3)]:
        triggered.add(127)
    if [(abs(x / z - 3.08) > 0.3)] != [(abs(x / z - 3.08) > 123)]:
        triggered.add(128)

    if [(abs((y / 10) / z - 0.354) > 0.03)] != [(abs((y / 10) / z - 0.354) > 3)]:
        triggered.add(129)
    if [(abs((y / 10) / z - 0.354) > 0.03)] != [(abs((y / 10) / z - 0.354) > 10.03)]:
        triggered.add(130)

    if [((x + y / 10 + z) < 292 or (x + y / 10 + z) > 304)] != [((x + y / 10 + z) < 292 or (x + y / 10 + z) > 3224)]:
        triggered.add(131)
    if [((x + y / 10 + z) < 292 or (x + y / 10 + z) > 304)] != [((x + y / 10 + z) < 292 or (x + y / 10 + z) > 1304)]:
        triggered.add(132)

    if [((x * y * z / 10000) < 29 or (x * y * z / 10000) > 35)] != [((x * y * z / 10000) < 29 or (x * y * z / 10000) > 135)]:
        triggered.add(133)
    if [((x * y * z / 10000) < 29 or (x * y * z / 10000) > 35)] != [((x * y * z / 10000) < 29 or (x * y * z / 10000) > 315)]:
        triggered.add(134)

    if [(x > 215)] != [(x > 1215)]:
        triggered.add(135)
    if [(x > 215)] != [(x > 2115)]:
        triggered.add(136)

    if [(x < 185)] != [(x < 1815)]:
        triggered.add(137)
    if [(x < 185)] != [(x < 1185)]:
        triggered.add(138)

    if [(y > 237)] != [(y > 2137)]:
        triggered.add(139)
    if [(y > 237)] != [(y > 2317)]:
        triggered.add(140)

    if [(y < 223)] != [(y < 2123)]:
        triggered.add(141)
    if [(y < 223)] != [(y < 2283)]:
        triggered.add(142)

    if [(z > 75)] != [(z > 751)]:
        triggered.add(143)
    if [(z > 75)] != [(z > 175)]:
        triggered.add(144)

    if [(z < 55)] != [(z < 515)]:
        triggered.add(145)
    if [(z < 55)] != [(z < 559)]:
        triggered.add(146)

    if [(x > 215 or x < 185) and (y > 237 or y < 223)] != [(x > 215 or x < 185) and (y > 237 or y < 2213)]:
        triggered.add(147)
    if [(x > 215 or x < 185) and (y > 237 or y < 223)] != [(x > 215 or x < 185) and (y > 237 or 200 < 223)]:
        triggered.add(148)

    if [(x > 215 or x < 185) and (z > 75 or z < 55)] != [(x > 215 or x < 185) and (z > 175 or z < 55)]:
        triggered.add(149)
    if [(x > 215 or x < 185) and (z > 75 or z < 55)] != [(x > 215 or x < 185) and (z > 75 or z < 551)]:
        triggered.add(150)

    if [(y > 237 or y < 223) and (z > 75 or z < 55)] != [(y > 237 or y < 223) and (z > 75 or z < 515)]:
        triggered.add(151)
    if [(y > 237 or y < 223) and (z > 75 or z < 55)] != [(y > 237 or y < 223) and (z > 175 or z < 55)]:
        triggered.add(152)

    if [(x > 215 or x < 185) and (y > 237 or y < 223) and (z > 75 or z < 55)] != [(x > 215 or x < 1185) and (y > 237 or y < 223) and (z > 75 or z < 55)]:
        triggered.add(153)
    if [(x > 215 or x < 185) and (y > 237 or y < 223) and (z > 75 or z < 55)] != [(x > 215 or x < 185) and (y > 2317 or y < 223) and (z > 75 or z < 55)]:
        triggered.add(154)

    if [(abs(x - 200) > 20) or (abs(y - 230) > 15) or (abs(z - 65) > 10)] != [(abs(x - 200) > 20) or (abs(y - 230) > 115) or (abs(z - 65) > 10)]:
        triggered.add(155)
    if [(abs(x - 200) > 20) or (abs(y - 230) > 15) or (abs(z - 65) > 10)] != [(abs(x - 200) > 20) or (abs(y - 230) > 415) or (abs(z - 65) > 10)]:
        triggered.add(156)

    if [(x > 220) or (y > 240) or (z > 80)] != [(x > 220) or (y > 240) or (z > 810)]:
        triggered.add(157)
    if [(x > 220) or (y > 240) or (z > 80)] != [(x > 220) or (y > 240) or (z > 911)]:
        triggered.add(158)
    if [(x < 175) or (y < 215) or (z < 45)] != [(x < 175) or (y < 215) or (z < 415)]:
        triggered.add(159)
    if [(x < 175) or (y < 215) or (z < 45)] != [(x < 175) or (y < 215) or (z < 145)]:
        triggered.add(160)
    return triggered

# ================================================================
# 3. 别名指向 section9（修正）
# ================================================================
execute_Tr = section9_hybrid_speed_torque_current

# ================================================================
# 4. Jaccard 相似度
# ================================================================
def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0

# ================================================================
# 5. 目标路径组（已为 section9 设计，保持不变）
# ================================================================
targetPaths = [
    {1, 2, 3, 4, 5, 6, 11, 12, 14, 16, 17, 18, 19, 21, 22, 23, 25, 26, 29, 30, 31, 34, 35, 36, 37, 38, 39, 40, 41, 42,
     47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100,
     101, 102, 103, 104, 108, 112, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137,
     138, 141, 142, 143, 144, 145, 146, 147, 148, 149, 157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 11, 12, 14, 15, 17, 18, 19, 23, 25, 26, 29, 30, 32, 34, 35, 36, 37, 38, 39, 40, 41, 42, 47, 48,
     49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102,
     103, 104, 108, 112, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 141,
     142, 143, 144, 145, 146, 147, 148, 149, 157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 11, 12, 14, 16, 17, 18, 19, 23, 25, 26, 29, 30, 31, 34, 35, 36, 37, 38, 39, 40, 41, 42, 47, 48,
     49, 50, 53, 54, 55, 56, 57, 58, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102, 103,
     104, 108, 111, 112, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 141,
     142, 143, 144, 145, 146, 147, 148, 149, 157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 14, 16, 17, 18, 19, 22, 25, 26, 29, 30, 31, 34, 35, 36, 37, 38, 39, 40, 41, 42, 47, 48, 49, 50,
     51, 52, 54, 55, 56, 57, 58, 63, 64, 67, 68, 69, 70, 73, 75, 76, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102,
     103, 104, 108, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 139, 140,
     141, 142, 143, 144, 145, 146, 149, 152, 154, 159, 160},

    {1, 2, 3, 4, 5, 6, 14, 16, 17, 18, 19, 21, 22, 23, 25, 26, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 47,
     48, 49, 50, 53, 54, 55, 56, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104,
     108, 111, 112, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 139, 140,
     141, 142, 143, 144, 145, 146, 149, 152, 154, 159, 160},

    {1, 2, 3, 4, 5, 6, 9, 10, 11, 12, 14, 17, 18, 19, 20, 21, 22, 25, 26, 27, 28, 29, 30, 32, 34, 35, 36, 37, 38, 39,
     40, 41, 42, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 85, 86, 87, 88, 95, 96,
     97, 98, 99, 100, 103, 104, 105, 108, 111, 112, 115, 116, 119, 120, 125, 126, 127, 128, 129, 130, 131, 132, 137,
     138, 141, 142, 143, 144, 145, 146, 157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 11, 12, 14, 17, 18, 19, 21, 22, 23, 25, 26, 29, 30, 31, 34, 35, 36, 37, 38, 39, 40, 41, 42, 47,
     48, 49, 50, 51, 52, 54, 55, 56, 57, 58, 59, 60, 63, 64, 67, 68, 69, 70, 73, 77, 78, 79, 80, 95, 96, 97, 98, 99,
     100, 103, 104, 108, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 141,
     142, 143, 144, 145, 146, 147, 148, 149, 159, 160},

    {1, 2, 3, 4, 5, 6, 9, 10, 14, 16, 17, 18, 19, 20, 21, 22, 25, 26, 29, 30, 31, 34, 35, 36, 37, 38, 39, 40, 47, 48,
     49, 50, 51, 52, 55, 56, 57, 58, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102, 103,
     104, 105, 108, 111, 112, 115, 116, 119, 120, 121, 122, 127, 128, 129, 130, 131, 132, 137, 138, 139, 140, 141, 142,
     143, 144, 145, 146, 152, 153, 157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 13, 16, 17, 18, 21, 22, 25, 26, 29, 30, 32, 34, 35, 36, 37, 38, 39, 40, 49, 50, 51,
     52, 55, 56, 57, 58, 63, 64, 69, 70, 77, 78, 79, 80, 87, 88, 93, 94, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104,
     105, 108, 111, 112, 115, 116, 119, 120, 121, 122, 127, 128, 129, 130, 131, 132, 137, 138, 139, 140, 141, 142, 143,
     144, 145, 146, 152, 153, 157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 16, 17, 18, 19, 20, 21, 22, 25, 26, 29, 30, 32, 33, 34, 35, 36, 37, 38, 49, 50, 51,
     52, 53, 54, 55, 56, 57, 58, 63, 64, 69, 70, 77, 78, 79, 80, 89, 90, 95, 96, 97, 98, 99, 100, 103, 104, 105, 108,
     111, 112, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 131, 132, 137, 138, 139, 140, 141, 142, 143,
     144, 145, 146, 152, 153, 157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 14, 16, 17, 18, 19, 21, 22, 23, 25, 26, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40, 47, 48, 49,
     50, 53, 54, 55, 56, 63, 64, 69, 70, 73, 74, 77, 78, 79, 80, 84, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 108,
     111, 112, 115, 116, 119, 120, 121, 122, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 139, 140, 141, 142, 143,
     144, 145, 146, 149, 152, 154, 159, 160},

    {1, 2, 3, 4, 5, 6, 11, 12, 14, 16, 17, 18, 19, 21, 22, 23, 24, 25, 26, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40,
     41, 42, 47, 48, 49, 50, 54, 55, 56, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104,
     108, 111, 112, 115, 116, 119, 120, 121, 122, 125, 126, 129, 130, 131, 132, 135, 136, 137, 138, 141, 142, 143, 144,
     145, 146, 147, 148, 149, 159, 160},

    {1, 2, 3, 4, 5, 6, 14, 16, 17, 18, 19, 21, 22, 23, 24, 25, 26, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42,
     43, 44, 47, 48, 49, 50, 54, 55, 56, 67, 68, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102, 103,
     104, 108, 112, 115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 131, 132, 135, 136, 137, 138, 139, 140, 141, 142,
     145, 146, 150, 151, 159, 160},

    {1, 2, 3, 4, 9, 10, 11, 12, 13, 15, 17, 18, 25, 26, 27, 28, 29, 30, 32, 34, 35, 36, 37, 38, 39, 40, 49, 50, 51, 52,
     55, 56, 57, 58, 63, 64, 69, 70, 77, 78, 79, 80, 81, 82, 85, 86, 87, 88, 95, 96, 97, 98, 99, 100, 101, 102, 103,
     104, 105, 108, 111, 112, 115, 116, 119, 120, 121, 122, 127, 128, 129, 130, 131, 132, 137, 138, 141, 142, 143, 144,
     145, 146, 157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 14, 16, 17, 18, 19, 22, 23, 24, 25, 26, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40, 47, 48, 49,
     50, 53, 54, 55, 56, 69, 70, 73, 74, 77, 78, 79, 80, 83, 84, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 108, 111,
     112, 115, 116, 119, 120, 121, 122, 129, 130, 131, 132, 135, 136, 137, 138, 139, 140, 141, 142, 143, 144, 145, 146,
     149, 152, 154, 159, 160},

    {1, 2, 3, 4, 5, 6, 9, 10, 11, 12, 13, 15, 17, 18, 23, 25, 26, 27, 28, 32, 33, 34, 35, 36, 37, 38, 49, 50, 55, 56,
     57, 58, 63, 64, 69, 70, 77, 78, 79, 80, 81, 82, 85, 86, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 105, 108, 111,
     112, 113, 114, 115, 116, 119, 120, 121, 122, 127, 128, 129, 130, 131, 132, 137, 138, 141, 142, 143, 144, 145, 146,
     157, 158, 159, 160},

    {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 13, 16, 17, 18, 22, 23, 24, 25, 26, 32, 33, 34, 35, 36, 49, 50, 55, 56, 63, 64, 69,
     70, 77, 78, 79, 80, 87, 88, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 108, 111, 112, 113, 114,
     115, 116, 119, 120, 121, 122, 125, 126, 127, 128, 129, 130, 137, 138, 139, 140, 141, 142, 143, 144, 145, 146, 152,
     153, 159, 160},

    {1, 2, 3, 4, 5, 6, 14, 16, 19, 22, 23, 24, 25, 26, 29, 30, 31, 33, 34, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47,
     48, 49, 50, 54, 55, 56, 67, 68, 69, 70, 77, 78, 79, 80, 97, 98, 99, 100, 103, 104, 108, 111, 112, 119, 120, 121,
     122, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 139, 140, 141, 142, 145, 146, 150, 151, 159, 160},

    {1, 2, 5, 6, 14, 16, 17, 18, 19, 22, 23, 24, 25, 26, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40, 47, 48, 49, 50, 53,
     54, 69, 70, 73, 74, 77, 78, 79, 80, 83, 84, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 108, 111, 112, 115, 116,
     119, 120, 121, 122, 131, 132, 135, 136, 137, 138, 139, 140, 141, 142, 145, 146, 150, 151, 155, 156, 159, 160},

    {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 13, 16, 17, 18, 22, 23, 24, 25, 26, 32, 33, 34, 35, 36, 53, 54, 55, 56, 63, 64, 69,
     70, 77, 78, 79, 80, 87, 88, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 107, 113, 114, 115, 116,
     118, 119, 120, 121, 122, 125, 126, 127, 128, 137, 138, 139, 140, 141, 142, 145, 146, 151, 155, 156, 159, 160},

    {1, 2, 5, 6, 9, 10, 14, 17, 18, 19, 20, 22, 23, 27, 28, 29, 30, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 49, 50, 54,
     55, 56, 57, 58, 63, 64, 69, 70, 73, 74, 79, 80, 87, 88, 95, 96, 97, 98, 99, 100, 103, 104, 108, 109, 111, 112, 115,
     116, 119, 120, 125, 126, 127, 128, 129, 130, 131, 132, 137, 138, 143, 144, 145, 146, 152, 153, 157, 158},

    {1, 2, 3, 4, 5, 6, 14, 16, 19, 22, 23, 24, 25, 26, 29, 30, 31, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49,
     50, 54, 55, 56, 57, 58, 61, 62, 65, 67, 68, 69, 70, 97, 98, 99, 100, 103, 104, 108, 109, 111, 112, 119, 120, 125,
     126, 127, 128, 129, 130, 131, 132, 135, 136, 137, 138, 139, 140, 141, 142, 145, 146, 150, 151, 159, 160},

    {1, 2, 3, 4, 5, 6, 14, 16, 19, 21, 22, 23, 24, 25, 26, 29, 30, 31, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48,
     49, 50, 54, 55, 56, 65, 69, 70, 97, 98, 99, 100, 103, 104, 108, 109, 111, 112, 119, 120, 121, 122, 125, 126, 127,
     128, 129, 130, 131, 132, 134, 135, 136, 137, 138, 141, 142, 145, 146, 147, 148, 150, 159, 160},

    {1, 2, 3, 4, 5, 6, 7, 8, 13, 16, 17, 18, 22, 23, 24, 25, 26, 32, 35, 36, 53, 54, 69, 70, 79, 80, 87, 88, 91, 92, 93,
     94, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 109, 111, 112, 115, 116, 117, 118, 119, 120, 121, 122, 125, 126,
     137, 138, 139, 140, 141, 142, 145, 146, 151, 155, 156, 159, 160},

    {1, 2, 3, 4, 5, 6, 9, 10, 11, 12, 14, 15, 17, 18, 19, 20, 23, 24, 25, 26, 32, 33, 34, 35, 36, 39, 40, 49, 50, 53,
     54, 69, 70, 79, 80, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 108, 109, 111, 112, 115, 116, 119, 120, 121, 122,
     123, 124, 129, 130, 137, 138, 141, 142, 145, 146, 159, 160},

    {3, 4, 5, 6, 14, 16, 19, 21, 22, 31, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 54, 56, 57, 58, 61, 62, 65, 66,
     67, 68, 69, 70, 79, 80, 99, 100, 103, 104, 108, 109, 110, 111, 112, 119, 120, 125, 126, 127, 128, 129, 130, 131,
     132, 134, 135, 136, 137, 138, 141, 142, 147, 148},

    {1, 2, 5, 6, 17, 18, 19, 22, 27, 28, 33, 34, 35, 36, 51, 52, 54, 56, 57, 58, 63, 64, 65, 69, 70, 71, 72, 89, 90, 95,
     96, 97, 98, 99, 100, 103, 104, 105, 106, 109, 112, 115, 116, 119, 120, 125, 126, 127, 128, 129, 130, 134, 143, 144,
     145, 146, 149, 152, 157, 158},

    {1, 2, 3, 4, 14, 19, 21, 22, 31, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 54, 57, 58, 59, 60, 61, 62, 65, 66, 75, 76,
     77, 78, 79, 80, 99, 100, 103, 104, 110, 111, 112, 119, 120, 125, 126, 127, 128, 129, 130, 133, 134, 135, 136, 137,
     138, 141, 142, 147, 148},
]

# ================================================================
# 6. 实验配置
# ================================================================
class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 2000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(targetPaths)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }

# ================================================================
# 7. 鲁棒性计算
# ================================================================
def compute_robustness(state, path):
    base = execute_Tr(*state)
    if not base:
        return 0.0
    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                    np.clip(state[0] + dx, STATE_MIN_X, STATE_MAX_X),
                    np.clip(state[1] + dy, STATE_MIN_Y, STATE_MAX_Y),
                    np.clip(state[2] + dz, STATE_MIN_Z, STATE_MAX_Z)
                ])
                n_trig = execute_Tr(*neighbor)
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0

# ================================================================
# 8. 候选样本生成（使用 randint，因为新范围为整数）
# ================================================================
def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = targetPaths[target_path_idx]
    samples = []
    attempts = 0
    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(STATE_MIN_X, STATE_MAX_X),
            random.randint(STATE_MIN_Y, STATE_MAX_Y),
            random.randint(STATE_MIN_Z, STATE_MAX_Z)
        ])
        triggered = execute_Tr(*state)
        if not triggered:
            continue
        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)
        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }
        samples.append(sample_data)
    return samples

def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample['robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)
        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]
    return selected_samples

def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = targetPaths[target_path_idx]
        samples = []
        attempts = 0
        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(STATE_MIN_X, STATE_MAX_X),
                random.randint(STATE_MIN_Y, STATE_MAX_Y),
                random.randint(STATE_MIN_Z, STATE_MAX_Z)
            ])
            triggered = execute_Tr(*state)
            if not triggered:
                continue
            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)
            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)
        random.shuffle(samples)
        return samples[:config.top_k_samples]
    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")
        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)

# ================================================================
# 9. 单次实验
# ================================================================
def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}
    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates
    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}
        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config, shared_candidates[path_idx])
            strategy_results[path_idx] = samples
        results[strategy_name] = strategy_results
    return results

# ================================================================
# 10. 结果分析
# ================================================================
def analyze_fitness_values(results, config):
    analysis_results = {}
    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []
        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])
        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }
        all_scores = []
        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])
        scores_array = np.array(all_scores)
        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)
        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })
        analysis_results[strategy_name] = analysis
    return analysis_results

def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []
    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)
    df = pd.DataFrame(df_data)
    return df, analysis_results

# ================================================================
# 11. 多轮实验
# ================================================================
def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []
    print(f"Starting {num_runs} experiments...")
    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")
        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)
        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)
        print(f"Experiment {run_idx + 1} completed.")
    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df

# ================================================================
# 12. 保存结果到 Excel
# ================================================================
def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "Raw Data"
    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')
    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width
    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []
    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]
        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)
    stats_df = pd.DataFrame(stats_data)
    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)
    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width
    wb.save(output_path)
    print(f"Results saved to: {output_path}")

# ================================================================
# 13. 主函数
# ================================================================
def main():
    results_df = run_multiple_experiments(num_runs=20)
    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")
    save_results_to_excel(results_df, output_path)
    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)
    return results_df, output_path

if __name__ == "__main__":
    results, output_file = main()