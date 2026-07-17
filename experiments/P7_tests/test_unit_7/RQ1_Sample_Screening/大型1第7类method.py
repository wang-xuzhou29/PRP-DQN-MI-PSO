import os
import random
import math
import numpy as np
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# --- 状态范围配置（按 section7 设计）---
STATE_MIN_X, STATE_MAX_X = 1, 80    # 温度/压力
STATE_MIN_Y, STATE_MAX_Y = 1, 120   # 电压/压力
STATE_MIN_Z, STATE_MAX_Z = 1, 15    # 流量/密度因子

def generate_input():
    return [
        random.randint(STATE_MIN_X, STATE_MAX_X),
        random.randint(STATE_MIN_Y, STATE_MAX_Y),
        random.randint(STATE_MIN_Z, STATE_MAX_Z)
    ]

# ========== 规则函数 section7 ==========
def section7_flow_pressure_density_hybrid(x, y, z):
    # 确保数值在范围内（安全防护）
    x = max(STATE_MIN_X, min(STATE_MAX_X, x))
    y = max(STATE_MIN_Y, min(STATE_MAX_Y, y))
    z = max(STATE_MIN_Z, min(STATE_MAX_Z, z))

    triggered = set()

    if (50 < x < 80) != (50 < x * 8 < 80):
        triggered.add(1)
    if (50 < x < 80) != (50 < 70 < 80):
        triggered.add(2)

    if (80 < y < 120) != (80 < y * 7 < 120):
        triggered.add(3)
    if (80 < y < 120) != (80 < 100 < 120):
        triggered.add(4)

    if (1.2 < z < 1.5) != (1.2 < z < 15):
        triggered.add(5)
    if (1.2 < z < 1.5) != (1.2 < z < 5):
        triggered.add(6)

    if (x > 60 and x < 70 and y > 90 and y < 110) != (x > 60 and x < 70 and y > 190 and y < 110):
        triggered.add(7)
    if (x > 60 and x < 70 and y > 90 and y < 110) != (x > 60 and x < 90 and y > 90 and y < 110):
        triggered.add(8)

    if (x > 55 and x < 75 and z > 1.25 and z < 1.45) != (x > 55 and x < 75 and z > 1.25 and z < 14.5):
        triggered.add(9)
    if (x > 55 and x < 75 and z > 1.25 and z < 1.45) != (x > 55 and x < 75 and z > 1.25 and z < 15):
        triggered.add(10)

    if (y > 85 and y < 115 and z > 1.25 and z < 1.45) != (y > 85 and y < 115 and z > 1.25 and z < 14.5):
        triggered.add(11)
    if (y > 85 and y < 115 and z > 1.25 and z < 1.45) != (y > 85 and y < 115 and z > 1.25 and z < 145):
        triggered.add(12)

    if (x > 63 and x < 67) != (x > 63 and x < 167):
        triggered.add(13)
    if (x > 63 and x < 67) != (x > 163 and x < 67):
        triggered.add(14)

    if (y > 98 and y < 102) != (y > 98 and y < 12.2):
        triggered.add(15)
    if (y > 98 and y < 102) != (y > 918 and y < 102):
        triggered.add(16)

    if (z > 1.33 and z < 1.37) != (z > 1.33 and z < 13.7):
        triggered.add(17)
    if (z > 1.33 and z < 1.37) != (z > 1.33 and 1 < 1.37):
        triggered.add(18)

    if (abs(x - 65) < 2.5) != (abs(x - 65) < 12.5):
        triggered.add(19)
    if (abs(x - 65) < 2.5) != (abs(x - 65) < 25):
        triggered.add(20)

    if (abs(y - 100) < 5) != (abs(y - 100) < 15):
        triggered.add(21)
    if (abs(y - 100) < 5) != (abs(y - 100) < 51):
        triggered.add(22)

    if (abs(z - 1.35) < 0.05) != (abs(z - 1.35) < 5):
        triggered.add(23)
    if (abs(z - 1.35) < 0.05) != (abs(z - 1.35) < 2.05):
        triggered.add(24)

    if (x > 55 and y > 85 and z > 1.25) != (x > 55 and y > 85 and z > 125):
        triggered.add(25)
    if (x > 55 and y > 85 and z > 1.25) != (x > 55 and y > 85 and 3 > 1.25):
        triggered.add(26)

    if (x < 75 and y < 115 and z < 1.45) != (x < 75 and y < 115 and z < 145):
        triggered.add(27)
    if (x < 75 and y < 115 and z < 1.45) != (x < 75 and y < 115 and z < 14.5):
        triggered.add(28)

    if ((x + y + z * 100) > 280 and (x + y + z * 100) < 320) != (
        (x + y + z * 100) > 480 and (x + y + z * 100) < 320):
        triggered.add(29)
    if ((x + y + z * 100) > 280 and (x + y + z * 100) < 320) != (
        (x + y + z * 100) > 280 and (x + y + z * 100) < 520):
        triggered.add(30)

    if (x * y > 5500 and x * y < 7500) != (x * y > 2500 and x * y < 7500):
        triggered.add(31)
    if (x * y > 5500 and x * y < 7500) != (x * y > 5500 and x * y < 4500):
        triggered.add(32)

    if (x * z > 70 and x * z < 110) != (x * z > 70 and x * z < 920):
        triggered.add(33)
    if (x * z > 70 and x * z < 110) != (x * z > 70 and x * z < 1110):
        triggered.add(34)

    if (y * z > 110 and y * z < 150) != (y * z > 1410 and y * z < 150):
        triggered.add(35)
    if (y * z > 110 and y * z < 150) != (y * z > 110 and y * z < 1550):
        triggered.add(36)

    if ((x + y + z * 100) / 3 > 93 and (x + y + z * 100) / 3 < 107) != (
        (x + y + z * 100) / 3 > 931 and (x + y + z * 100) / 3 < 107):
        triggered.add(37)
    if ((x + y + z * 100) / 3 > 93 and (x + y + z * 100) / 3 < 107) != (
        (x + y + z * 100) / 3 > 93 and (x + y + z * 200) / 3 < 107):
        triggered.add(38)

    if (math.sqrt(x ** 2 + y ** 2 + (z * 100) ** 2) > 150) != (math.sqrt(x ** 2 + y ** 2 + (z * 100) ** 2) > 1510):
        triggered.add(39)
    if (math.sqrt(x ** 2 + y ** 2 + (z * 100) ** 2) > 150) != (math.sqrt(x ** 2 + y ** 2 + (z * 100) ** 2) > 120):
        triggered.add(40)

    if (x / y > 0.55 and x / y < 0.85) != (x / y > 0.55 and x / y < 85):
        triggered.add(41)
    if (x / y > 0.55 and x / y < 0.85) != (x / y > 0.55 and x / y < 8.5):
        triggered.add(42)

    if (x / z > 40 and x / z < 60) != (x / z > 40 and x / z < 610):
        triggered.add(43)
    if (x / z > 40 and x / z < 60) != (x / z > 401 and x / z < 60):
        triggered.add(44)

    if (y / z > 60 and y / z < 90) != (y / z > 60 and y / z < 910):
        triggered.add(45)
    if (y / z > 60 and y / z < 90) != (y / z > 160 and y / z < 90):
        triggered.add(46)

    if ((x - 50) + (y - 80) > 20 and (x - 50) + (y - 80) < 40) != (
        (x - 50) + (y - 80) > 20 and (x - 50) + (y - 80) < 410):
        triggered.add(47)
    if ((x - 50) + (y - 80) > 20 and (x - 50) + (y - 80) < 40) != (
        (x - 50) + (y - 80) > 20 and (x - 50) + (y - 80) < 140):
        triggered.add(48)

    if ((x - 50) + (z - 1.2) * 100 > 25 and (x - 50) + (z - 1.2) * 100 < 45) != (
        (x - 50) + (z - 1.2) * 100 > 25 and (x - 50) + (z - 1.2) * 100 < 145):
        triggered.add(49)
    if ((x - 50) + (z - 1.2) * 100 > 25 and (x - 50) + (z - 1.2) * 100 < 45) != (
        (x - 50) + (z - 1.2) * 100 > 25 and (x - 50) + (z - 1.2) * 100 < 451):
        triggered.add(50)

    if ((y - 80) + (z - 1.2) * 100 > 35 and (y - 80) + (z - 1.2) * 100 < 55) != (
        (y - 80) + (z - 1.2) * 100 > 35 and (y - 80) + (z - 1.2) * 100 < 155):
        triggered.add(51)
    if ((y - 80) + (z - 1.2) * 100 > 35 and (y - 80) + (z - 1.2) * 100 < 55) != (
        (y - 80) + (z - 1.2) * 100 > 35 and (y - 80) + (z - 1.2) * 100 < 515):
        triggered.add(52)

    if (abs((x - 65) - (y - 100) * 0.65) < 8) != (abs((x - 65) - (y - 100) * 0.65) < 18):
        triggered.add(53)
    if (abs((x - 65) - (y - 100) * 0.65) < 8) != (abs((x - 65) - (y - 100) * 0.65) < 81):
        triggered.add(54)

    if (abs((x - 65) - (z - 1.35) * 50) < 10) != (abs((x - 65) - (z - 1.35) * 50) < 110):
        triggered.add(55)
    if (abs((x - 65) - (z - 1.35) * 50) < 10) != (abs((x - 65) - (z - 1.35) * 50) < 101):
        triggered.add(56)

    if (abs((y - 100) - (z - 1.35) * 100) < 12) != (abs((y - 100) - (z - 1.35) * 100) < 112):
        triggered.add(57)
    if (abs((y - 100) - (z - 1.35) * 100) < 12) != (abs((y - 100) - (z - 1.35) * 100) < 132):
        triggered.add(58)

    if (x / (y + 20) > 0.5 and x / (y + 20) < 0.8) != (x / (y + 20) > 0.5 and x / (y + 20) < 8):
        triggered.add(59)
    if (x / (y + 20) > 0.5 and x / (y + 20) < 0.8) != (x / (y + 20) > 0.5 and x / (y + 20) < 48):
        triggered.add(60)

    if (z / (x / 50) > 0.9 and z / (x / 50) < 1.4) != (z / (x / 50) > 0.9 and z / (x / 50) < 14.4):
        triggered.add(61)
    if (z / (x / 50) > 0.9 and z / (x / 50) < 1.4) != (z / (x / 50) > 0.9 and z / (x / 50) < 133.4):
        triggered.add(62)

    if (x * y * z > 7000 and x * y * z < 11000) != (x * y * z > 7000 and x * y * 2 < 11000):
        triggered.add(63)
    if (x * y * z > 7000 and x * y * z < 11000) != (x * y * z > 7000 and x * y * 54 < 11000):
        triggered.add(64)

    if (x * 0.6 + y * 0.3 + z * 10 > 65 and x * 0.6 + y * 0.3 + z * 10 < 75) != (
        x * 0.6 + y * 0.3 + z * 10 > 645 and x * 0.6 + y * 0.3 + z * 10 < 75):
        triggered.add(65)
    if (x * 0.6 + y * 0.3 + z * 10 > 65 and x * 0.6 + y * 0.3 + z * 10 < 75) != (
        x * 0.6 + y * 40.3 + z * 10 > 65 and x * 0.6 + y * 0.3 + z * 10 < 75):
        triggered.add(66)

    if ((x / 65) ** 0.5 * (y / 100) ** 0.3 * (z / 1.35) ** 0.2 > 0.92) != (
        (x / 65) ** 0.5 * (y / 100) ** 0.3 * (z / 1.35) ** 0.2 > 392):
        triggered.add(67)
    if ((x / 65) ** 0.5 * (y / 100) ** 0.3 * (z / 1.35) ** 0.2 > 0.92) != (
        (x / 65) ** 0.5 * (y / 100) ** 0.3 * (z / 1.35) ** 0.2 > 982):
        triggered.add(68)

    if ((x - 65) * (y - 100) > -150 and (x - 65) * (y - 100) < 150) != (
        (x - 65) * (y - 100) > -150 and (x - 65) * (y - 100) < 1250):
        triggered.add(69)
    if ((x - 65) * (y - 100) > -150 and (x - 65) * (y - 100) < 150) != (
        (x - 65) * (y - 100) > -150 and (x - 65) * (y - 100) < 1050):
        triggered.add(70)

    if ((x - 65) * (z - 1.35) > -8 and (x - 65) * (z - 1.35) < 8) != (
        (x - 65) * (z - 1.35) > -8 and (x - 65) * (z - 1.35) < 18):
        triggered.add(71)
    if ((x - 65) * (z - 1.35) > -8 and (x - 65) * (z - 1.35) < 8) != (
        (x - 65) * (z - 1.35) > -8 and (x - 65) * (z - 1.35) < 448):
        triggered.add(72)

    if (abs(x / 65 + y / 100 + z / 1.35 - 3) < 0.25) != (abs(x / 65 + y / 100 + z / 1.35 - 3) < 25):
        triggered.add(73)
    if (abs(x / 65 + y / 100 + z / 1.35 - 3) < 0.25) != (abs(x / 65 + y / 100 + z / 1.35 - 3) < 40.25):
        triggered.add(74)

    if (abs((x / 65) * (y / 100) * (z / 1.35) - 1) < 0.18) != (abs((x / 65) * (y / 100) * (z / 1.35) - 1) < 40.18):
        triggered.add(75)
    if (abs((x / 65) * (y / 100) * (z / 1.35) - 1) < 0.18) != (abs((x / 65) * (y / 100) * (z / 1.35) - 1) < 18):
        triggered.add(76)

    if (x > 62 and x < 68 and y > 97 and y < 103 and z > 1.32 and z < 1.38) != (
        x > 62 and x < 68 and y > 97 and y < 103 and z > 1.32 and z < 138):
        triggered.add(77)
    if (x > 62 and x < 68 and y > 97 and y < 103 and z > 1.32 and z < 1.38) != (
        x > 62 and x < 68 and y > 97 and y < 103 and z > 1.32 and z < 228):
        triggered.add(78)

    if (abs(x / y - 0.65) < 0.08 and abs(z - 1.35) < 0.03) != (abs(x / y - 0.65) < 0.08 and abs(z - 1.35) < 3):
        triggered.add(79)
    if (abs(x / y - 0.65) < 0.08 and abs(z - 1.35) < 0.03) != (abs(x / y - 0.65) < 0.08 and abs(z - 1.35) < 40.03):
        triggered.add(80)

    if (x * y > 6300 and x * y < 6700 and z > 1.33 and z < 1.37) != (
        x * y > 6300 and x * y < 6700 and z > 1.33 and z < 41.37):
        triggered.add(81)
    if (x * y > 6300 and x * y < 6700 and z > 1.33 and z < 1.37) != (
        x * y > 6300 and x * y < 6700 and z > 1.33 and z < 91.37):
        triggered.add(82)

    if ((x + y) / 2 > 80 and (x + y) / 2 < 86 and z > 1.32 and z < 1.38) != (
        (x + y) / 2 > 80 and (x + y) / 2 < 86 and z > 1.32 and z < 138):
        triggered.add(83)
    if ((x + y) / 2 > 80 and (x + y) / 2 < 86 and z > 1.32 and z < 1.38) != (
        (x + y) / 2 > 80 and (x + y) / 2 < 86 and z > 1.32 and z < 19.38):
        triggered.add(84)

    if (abs(x - y) < 40 and z > 1.32 and z < 1.38) != (abs(x - y) < 40 and z > 1.32 and z < 138):
        triggered.add(85)
    if (abs(x - y) < 40 and z > 1.32 and z < 1.38) != (abs(x - y) < 40 and z > 1.32 and z < 13.8):
        triggered.add(86)

    if (math.sqrt((x - 65) ** 2 + (y - 100) ** 2) < 5 and abs(z - 1.35) < 0.03) != (
        math.sqrt((x - 65) ** 2 + (y - 100) ** 2) < 5 and abs(z - 1.35) < 13):
        triggered.add(87)
    if (math.sqrt((x - 65) ** 2 + (y - 100) ** 2) < 5 and abs(z - 1.35) < 0.03) != (
        math.sqrt((x - 65) ** 2 + (y - 100) ** 2) < 5 and abs(z - 1.35) < 8):
        triggered.add(88)

    if (x / 65 > 0.975 and x / 65 < 1.025 and y / 100 > 0.97 and y / 100 < 1.03) != (
        x / 65 > 0.975 and x / 65 < 1.025 and y / 100 > 10.97 and y / 100 < 1.03):
        triggered.add(89)
    if (x / 65 > 0.975 and x / 65 < 1.025 and y / 100 > 0.97 and y / 100 < 1.03) != (
        x / 65 > 0.975 and x / 65 < 1.025 and y / 100 > 40.97 and y / 100 < 1.03):
        triggered.add(90)

    if (z / 1.35 > 0.978 and z / 1.35 < 1.022) != (z / 1.35 > 0.978 and z / 1.35 < 41.022):
        triggered.add(91)
    if (z / 1.35 > 0.978 and z / 1.35 < 1.022) != (z / 1.35 > 0.978 and z / 1.35 < 622):
        triggered.add(92)

    if ((x / 65 + y / 100 + z / 1.35) / 3 > 0.975 and (x / 65 + y / 100 + z / 1.35) / 3 < 1.025) != (
        (x / 65 + y / 100 + z / 1.35) / 3 > 9.75 and (x / 65 + y / 100 + z / 1.35) / 3 < 1.025):
        triggered.add(93)
    if ((x / 65 + y / 100 + z / 1.35) / 3 > 0.975 and (x / 65 + y / 100 + z / 1.35) / 3 < 1.025) != (
        (x / 65 + y / 100 + z / 1.35) / 3 > 0.975 and (x / 65 + y / 100 + z / 1.35) / 3 < 10.25):
        triggered.add(94)

    if (max(abs(x / 65 - 1), abs(y / 100 - 1), abs(z / 1.35 - 1)) < 0.025) != (
        max(abs(x / 65 - 1), abs(y / 100 - 1), abs(z / 1.35 - 1)) < 40.025):
        triggered.add(95)
    if (max(abs(x / 65 - 1), abs(y / 100 - 1), abs(z / 1.35 - 1)) < 0.025) != (
        max(abs(x / 65 - 1), abs(y / 100 - 1), abs(z / 1.35 - 1)) < 10.025):
        triggered.add(96)

    if (min(x / 65, y / 100, z / 1.35) > 0.975) != (min(x / 65, y / 100, z / 1.35) > 9.75):
        triggered.add(97)
    if (min(x / 65, y / 100, z / 1.35) > 0.975) != (min(x / 65, y / 100, z / 1.35) > 19.75):
        triggered.add(98)

    if (max(x / 65, y / 100, z / 1.35) < 1.025) != (max(x / 65, y / 100, z / 1.35) < 11.025):
        triggered.add(99)
    if (max(x / 65, y / 100, z / 1.35) < 1.025) != (max(x / 65, y / 100, z / 1.35) < 41.025):
        triggered.add(100)

    if (abs(max(x, y, z * 100) - min(x, y, z * 100)) < 40) != (abs(max(x, y, z * 100) - min(x, y, z * 100)) < 140):
        triggered.add(101)
    if (abs(max(x, y, z * 100) - min(x, y, z * 100)) < 40) != (abs(max(x, y, z * 100) - min(x, y, z * 100)) < 1140):
        triggered.add(102)

    if ((x + y + z * 100) > 298 and (x + y + z * 100) < 302) != (
        (x + y + z * 100) > 298 and (x + y + z * 100) < 3102):
        triggered.add(103)
    if ((x + y + z * 100) > 298 and (x + y + z * 100) < 302) != (
        (x + y + z * 100) > 298 and (x + y + z * 100) < 352):
        triggered.add(104)

    if (x * y * z > 8700 and x * y * z < 9300) != (x * y * z > 8700 and x * y * 4 < 9300):
        triggered.add(105)
    if (x * y * z > 8700 and x * y * z < 9300) != (x * y * z > 8700 and x * y * z < 5300):
        triggered.add(106)

    if (abs((x + y + z * 100) / 3 - 100) < 2) != (abs((x + y + z * 100) / 3 - 100) < 332):
        triggered.add(107)
    if (abs((x + y + z * 100) / 3 - 100) < 2) != (abs((x + y + z * 100) / 3 - 100) < 122):
        triggered.add(108)

    if (z > 1.345 and z < 1.355) != (z > 1.345 and z < 1355):
        triggered.add(109)
    if (z > 1.345 and z < 1.355) != (z > 1.345 and z < 355):
        triggered.add(110)

    if (x < 52 or y < 85 or z < 1.22) != (x < 52 or y < 85 or z < 122):
        triggered.add(111)
    if (x < 52 or y < 85 or z < 1.22) != (x < 52 or y < 855 or z < 1.22):
        triggered.add(112)

    if (x > 78 or y > 115 or z > 1.48) != (x > 78 or y > 115 or z > 148):
        triggered.add(113)
    if (x > 78 or y > 115 or z > 1.48) != (x > 78 or y > 115 or z * 8 > 1.48):
        triggered.add(114)

    if (abs(x / y - 0.65) > 0.12) != (abs(x / y - 0.65) > 12):
        triggered.add(115)
    if (abs(x / y - 0.65) > 0.12) != (abs(x / y - 0.65) > 20.12):
        triggered.add(116)

    if (abs(x / z - 48) > 6) != (abs(x / z - 48) > 16):
        triggered.add(117)
    if (abs(x / z - 48) > 6) != (abs(x / z - 48) > 36):
        triggered.add(118)

    if (abs(y / z - 74) > 8) != (abs(y / z - 74) > 228):
        triggered.add(119)
    if (abs(y / z - 74) > 8) != (abs(y / z - 74) > 82):
        triggered.add(120)

    if ((x + y + z * 100) < 285 or (x + y + z * 100) > 315) != (
        (x + y + z * 100) < 285 or (x + y + z * 100) > 3165):
        triggered.add(121)
    if ((x + y + z * 100) < 285 or (x + y + z * 100) > 315) != (
        (x + y + z * 100) < 2835 or (x + y + z * 100) > 315):
        triggered.add(122)

    if (x * y * z < 8000 or x * y * z > 10000) != (x * y * z < 8000 or x * y * z > 1000):
        triggered.add(123)
    if (x * y * z < 8000 or x * y * z > 10000) != (x * y * z < 8000 or x * y * z > 93000):
        triggered.add(124)

    if (x < 55 and y < 90) != (x < 55 and y < 290):
        triggered.add(125)
    if (x < 55 and y < 90) != (x < 55 and 80 < 90):
        triggered.add(126)

    if (x > 75 and y > 110) != (x > 75 and 200 > 110):
        triggered.add(127)
    if (x > 75 and y > 110) != (x > 75 and 500 > 110):
        triggered.add(128)

    if (x < 55 and z < 1.25) != (x < 55 and z < 125):
        triggered.add(129)
    if (x < 55 and z < 1.25) != (x < 55 and z < 12.5):
        triggered.add(130)

    if (x > 75 and z > 1.45) != (x > 75 and z > 145):
        triggered.add(131)
    if (x > 75 and z > 1.45) != (x > 75 and 6 > 1.45):
        triggered.add(132)

    if (y < 90 and z < 1.25) != (y < 90 and 1 < 1.25):
        triggered.add(133)
    if (y < 90 and z < 1.25) != (y < 90 and z < 125):
        triggered.add(134)

    if (y < 75 or y > 125) != (y < 75 or y * 8 > 125):
        triggered.add(135)
    if (y < 75 or y > 125) != (y < 75 or y * 10 > 125):
        triggered.add(136)

    if (x < 45 and y < 80 and z < 1.2) != (x < 45 and y < 80 and z < 12):
        triggered.add(137)
    if (x < 45 and y < 80 and z < 1.2) != (x < 45 and y < 80 and z < 11.2):
        triggered.add(138)

    if (abs((x * y * z) / 9000 - 1) > 0.22) != (abs((x * y * z) / 9000 - 1) > 9.22):
        triggered.add(139)
    if (abs((x * y * z) / 9000 - 1) > 0.22) != (abs((x * y * z) / 9000 - 1) > 22):
        triggered.add(140)

    return triggered

# ========== 关键修正：别名指向 section7 ==========
execute_Tr = section7_flow_pressure_density_hybrid

# ========== Jaccard 相似度 ==========
def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0

# === 目标路径组（为 section7 设计） ===
targetPaths = [
    {1, 3, 5, 6, 7, 9, 10, 11, 12, 13, 17, 18, 19, 20, 23, 24, 25, 27, 28, 30, 32, 33, 34, 36, 39, 47, 48, 49, 50, 51,
     52, 55, 56, 57, 58, 61, 62, 67, 68, 73, 74, 75, 76, 79, 80, 83, 84, 85, 86, 87, 88, 91, 92, 94, 95, 96, 97, 98, 99,
     100, 101, 102, 103, 107, 108, 109, 110, 111, 112, 113, 117, 118, 119, 120, 121, 124, 135, 136, 139, 140},

    {1, 3, 5, 6, 7, 9, 10, 11, 12, 13, 15, 16, 17, 18, 19, 20, 23, 24, 25, 27, 28, 30, 32, 33, 34, 36, 39, 49, 50, 51,
     52, 55, 56, 57, 58, 61, 62, 67, 68, 73, 74, 75, 76, 79, 80, 83, 84, 85, 86, 87, 88, 91, 92, 94, 95, 96, 97, 98, 99,
     100, 101, 102, 103, 107, 108, 109, 110, 111, 112, 113, 117, 118, 119, 120, 121, 124, 135, 136, 139, 140},

    {1, 3, 5, 6, 9, 10, 11, 12, 13, 17, 18, 19, 20, 21, 22, 23, 24, 25, 27, 28, 30, 32, 33, 34, 36, 39, 49, 50, 51, 52,
     53, 54, 55, 56, 57, 58, 61, 62, 67, 68, 73, 74, 75, 76, 81, 82, 83, 84, 85, 86, 91, 92, 94, 95, 96, 99, 100, 101,
     102, 103, 107, 108, 109, 110, 111, 112, 113, 115, 116, 117, 118, 119, 120, 121, 124, 135, 136, 139, 140},

    {2, 3, 5, 6, 11, 12, 13, 17, 18, 20, 21, 22, 23, 24, 25, 30, 32, 33, 34, 36, 39, 41, 42, 49, 50, 51, 52, 54, 55, 56,
     57, 58, 67, 68, 71, 72, 73, 74, 75, 76, 83, 84, 85, 86, 91, 92, 94, 95, 96, 99, 100, 101, 102, 103, 107, 108, 109,
     110, 111, 112, 115, 116, 117, 118, 119, 120, 121, 124, 127, 128, 131, 133, 134, 135, 136, 139, 140},

    {1, 3, 5, 6, 8, 9, 10, 11, 12, 13, 17, 18, 19, 20, 21, 22, 23, 24, 25, 27, 28, 30, 32, 33, 34, 36, 39, 49, 50, 51,
     52, 53, 54, 55, 56, 57, 58, 67, 68, 73, 74, 75, 76, 81, 82, 83, 84, 85, 86, 91, 92, 94, 95, 96, 99, 100, 101, 102,
     103, 107, 108, 109, 110, 111, 112, 113, 115, 116, 117, 118, 119, 120, 121, 124, 135, 136, 139, 140},

    {1, 3, 5, 6, 7, 9, 10, 11, 12, 15, 16, 17, 18, 23, 24, 25, 27, 28, 30, 32, 33, 34, 36, 39, 49, 50, 51, 52, 55, 56,
     57, 58, 61, 62, 67, 68, 73, 74, 75, 76, 77, 78, 79, 80, 83, 84, 85, 86, 87, 88, 91, 92, 94, 95, 96, 99, 100, 101,
     102, 103, 107, 108, 109, 110, 111, 112, 113, 118, 119, 120, 121, 124, 135, 136, 139, 140},

    {1, 3, 5, 6, 9, 10, 11, 12, 14, 17, 18, 21, 22, 23, 24, 25, 27, 28, 30, 32, 33, 34, 36, 39, 47, 48, 49, 50, 51, 52,
     53, 54, 55, 56, 57, 58, 61, 62, 67, 68, 73, 74, 75, 76, 79, 80, 91, 92, 94, 95, 96, 97, 98, 99, 100, 101, 102, 103,
     107, 108, 109, 110, 111, 112, 113, 117, 118, 119, 120, 121, 124, 135, 136, 139, 140},

    {1, 3, 5, 6, 9, 10, 13, 17, 18, 19, 20, 22, 23, 24, 27, 28, 30, 32, 33, 34, 36, 39, 41, 42, 49, 50, 51, 52, 53, 54,
     55, 56, 57, 58, 61, 62, 67, 68, 73, 74, 75, 76, 85, 86, 91, 92, 94, 95, 96, 99, 100, 101, 102, 103, 104, 107, 108,
     109, 110, 113, 115, 116, 117, 118, 119, 120, 121, 124, 133, 134, 135, 136, 139, 140},

    {1, 4, 5, 6, 13, 17, 18, 20, 22, 23, 24, 30, 32, 33, 34, 36, 39, 41, 42, 49, 50, 51, 52, 54, 55, 56, 57, 58, 59, 60,
     67, 68, 71, 72, 73, 74, 75, 76, 85, 86, 91, 92, 94, 95, 96, 99, 100, 101, 102, 103, 107, 108, 109, 110, 113, 115,
     116, 117, 118, 119, 120, 121, 124, 127, 128, 131, 133, 134, 135, 136, 139, 140},

    {1, 3, 5, 6, 11, 12, 13, 17, 18, 20, 21, 22, 23, 24, 25, 30, 33, 34, 36, 39, 47, 48, 49, 50, 51, 52, 55, 56, 57, 58,
     67, 68, 69, 70, 71, 72, 73, 74, 75, 76, 79, 80, 85, 86, 91, 92, 94, 95, 96, 97, 98, 99, 100, 101, 102, 103, 107,
     108, 109, 110, 111, 112, 117, 118, 119, 120, 121, 124, 131, 135, 136, 139, 140},

    {1, 4, 5, 6, 9, 10, 13, 17, 18, 19, 20, 22, 23, 24, 27, 28, 30, 31, 33, 34, 36, 39, 41, 42, 49, 50, 51, 52, 54, 55,
     56, 57, 58, 61, 62, 64, 67, 68, 73, 74, 75, 76, 85, 86, 91, 92, 94, 95, 96, 99, 100, 101, 102, 103, 104, 107, 108,
     109, 110, 113, 115, 116, 117, 118, 119, 120, 121, 124, 133, 134, 135, 136},
    {1, 3, 5, 7, 9, 10, 11, 12, 14, 15, 16, 17, 18, 23, 25, 27, 28, 32, 33, 34, 36, 39, 50, 52, 61, 62, 67, 68, 73, 74,
     75, 76, 77, 78, 80, 81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 94, 95, 96, 97, 98, 99, 100, 102, 103, 107,
     108, 109, 110, 111, 112, 113, 118, 119, 120, 121, 124, 135, 136, 139, 140},

    {1, 4, 5, 6, 13, 17, 18, 19, 20, 22, 23, 24, 30, 32, 33, 34, 35, 39, 41, 42, 49, 50, 51, 52, 54, 55, 56, 57, 58, 59,
     60, 67, 68, 73, 74, 75, 76, 85, 86, 91, 92, 94, 95, 96, 99, 100, 101, 102, 103, 104, 107, 108, 109, 110, 113, 115,
     116, 117, 118, 119, 120, 121, 124, 127, 128, 131, 133, 134, 139, 140},
    {1, 3, 5, 6, 11, 12, 17, 18, 20, 21, 22, 23, 24, 27, 28, 30, 32, 36, 39, 49, 50, 51, 52, 54, 55, 56, 57, 58, 61, 62,
     67, 68, 73, 74, 75, 76, 83, 84, 91, 92, 94, 95, 96, 99, 100, 102, 103, 107, 108, 109, 110, 111, 112, 113, 115, 116,
     118, 119, 120, 121, 124, 125, 126, 129, 130, 135, 136, 139, 140},

    {1, 3, 5, 6, 11, 12, 17, 18, 19, 20, 21, 22, 23, 24, 27, 28, 30, 31, 36, 39, 49, 50, 51, 52, 55, 56, 57, 58, 61, 62,
     64, 67, 68, 69, 70, 79, 80, 85, 86, 91, 92, 94, 95, 96, 99, 100, 102, 103, 104, 105, 106, 107, 108, 109, 110, 111,
     112, 113, 118, 119, 120, 121, 123, 129, 130, 133, 134, 135, 136},

    {1, 4, 5, 6, 9, 10, 13, 17, 18, 19, 20, 22, 23, 24, 27, 28, 30, 31, 33, 34, 36, 39, 41, 42, 50, 51, 52, 54, 55, 56,
     59, 60, 61, 62, 63, 67, 68, 71, 72, 73, 74, 75, 76, 85, 86, 91, 92, 94, 95, 96, 99, 100, 102, 103, 107, 108, 109,
     110, 113, 115, 116, 118, 119, 120, 121, 124, 133, 134, 139, 140},

    {1, 4, 5, 6, 17, 18, 19, 20, 22, 23, 24, 27, 28, 29, 31, 35, 37, 38, 39, 41, 42, 49, 50, 51, 52, 53, 54, 55, 56, 57,
     58, 61, 62, 65, 69, 70, 75, 76, 85, 86, 91, 92, 93, 95, 96, 99, 100, 102, 103, 104, 107, 108, 109, 110, 113, 115,
     116, 118, 119, 120, 121, 129, 130, 133, 134, 139, 140},

    {2, 4, 5, 6, 17, 18, 22, 23, 24, 27, 28, 29, 31, 36, 37, 38, 39, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 61, 62, 66,
     69, 70, 75, 76, 85, 86, 91, 92, 95, 96, 99, 100, 102, 103, 104, 107, 108, 109, 110, 113, 115, 116, 118, 119, 120,
     122, 129, 130, 133, 134, 135, 136, 137, 138, 139, 140},

    {1, 4, 13, 19, 20, 22, 23, 24, 26, 29, 35, 37, 38, 39, 43, 45, 47, 48, 55, 56, 57, 58, 63, 64, 67, 68, 69, 70, 79,
     80, 94, 95, 96, 99, 100, 101, 102, 105, 106, 107, 108, 118, 119, 120, 122, 123, 135, 136},

    {1, 3, 13, 20, 21, 22, 23, 24, 26, 29, 35, 37, 38, 39, 43, 45, 47, 48, 55, 56, 57, 58, 63, 64, 67, 68, 69, 70, 79,
     80, 93, 95, 96, 99, 100, 105, 106, 107, 108, 118, 119, 120, 122, 123, 132, 135, 136},

    {1, 3, 8, 13, 19, 20, 21, 22, 23, 24, 26, 29, 37, 38, 39, 43, 45, 47, 48, 55, 56, 57, 58, 63, 64, 67, 68, 79, 80,
     93, 95, 96, 99, 100, 107, 108, 114, 118, 119, 120, 123, 127, 128, 132, 135, 136},

    {2, 3, 21, 22, 23, 24, 31, 40, 45, 54, 55, 56, 57, 58, 61, 62, 66, 71, 72, 73, 74, 75, 76, 95, 96, 99, 100, 101,
     102, 107, 108, 114, 115, 116, 117, 118, 119, 120, 125, 126, 135, 136, 139, 140},

    {2, 3, 20, 21, 22, 23, 24, 31, 40, 44, 45, 53, 54, 57, 58, 66, 69, 70, 71, 72, 73, 74, 75, 76, 95, 96, 101, 102,
     107, 108, 114, 115, 116, 117, 118, 119, 120, 125, 126, 135, 136, 139, 140},

    {2, 3, 21, 22, 23, 24, 31, 40, 46, 54, 55, 56, 57, 58, 61, 62, 66, 69, 70, 71, 72, 73, 74, 75, 76, 95, 96, 101, 102,
     107, 108, 114, 115, 116, 117, 118, 119, 120, 135, 136, 139, 140},
]

# ========== 实验配置 ==========
class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 2000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(targetPaths)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }

def compute_robustness(state, path):
    base = execute_Tr(*state)
    if not base:
        return 0.0
    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                    np.clip(state[0] + dx, STATE_MIN_X, STATE_MAX_X),
                    np.clip(state[1] + dy, STATE_MIN_Y, STATE_MAX_Y),
                    np.clip(state[2] + dz, STATE_MIN_Z, STATE_MAX_Z)
                ])
                n_trig = execute_Tr(*neighbor)
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0

def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = targetPaths[target_path_idx]
    samples = []
    attempts = 0
    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(STATE_MIN_X, STATE_MAX_X),
            random.randint(STATE_MIN_Y, STATE_MAX_Y),
            random.randint(STATE_MIN_Z, STATE_MAX_Z)
        ])
        triggered = execute_Tr(*state)
        if not triggered:
            continue
        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)
        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }
        samples.append(sample_data)
    return samples

def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample['robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)
        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]
    return selected_samples

def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = targetPaths[target_path_idx]
        samples = []
        attempts = 0
        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(STATE_MIN_X, STATE_MAX_X),
                random.randint(STATE_MIN_Y, STATE_MAX_Y),
                random.randint(STATE_MIN_Z, STATE_MAX_Z)
            ])
            triggered = execute_Tr(*state)
            if not triggered:
                continue
            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)
            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)
        random.shuffle(samples)
        return samples[:config.top_k_samples]
    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")
        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)

def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}
    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates
    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}
        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config, shared_candidates[path_idx])
            strategy_results[path_idx] = samples
        results[strategy_name] = strategy_results
    return results

def analyze_fitness_values(results, config):
    analysis_results = {}
    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []
        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])
        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }
        all_scores = []
        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])
        scores_array = np.array(all_scores)
        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)
        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })
        analysis_results[strategy_name] = analysis
    return analysis_results

def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []
    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)
    df = pd.DataFrame(df_data)
    return df, analysis_results

def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []
    print(f"Starting {num_runs} experiments...")
    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")
        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)
        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)
        print(f"Experiment {run_idx + 1} completed.")
    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df

def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "Raw Data"
    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')
    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width
    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []
    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]
        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)
    stats_df = pd.DataFrame(stats_data)
    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)
    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width
    wb.save(output_path)
    print(f"Results saved to: {output_path}")

def main():
    results_df = run_multiple_experiments(num_runs=20)
    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")
    save_results_to_excel(results_df, output_path)
    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)
    return results_df, output_path

if __name__ == "__main__":
    results, output_file = main()