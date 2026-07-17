import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

def execute_validation_rules(x, y, z):
    # 确保数值在范围内（安全防护）
    x = max(STATE_MIN_X, min(STATE_MAX_X, x))
    y = max(STATE_MIN_Y, min(STATE_MAX_Y, y))
    z = max(STATE_MIN_Z, min(STATE_MAX_Z, z))

    triggered = set()

    if (50 < x < 80) != (50 < x * 8 < 80):
        triggered.add(1)
    if (50 < x < 80) != (50 < 70 < 80):
        triggered.add(2)

    if (80 < y < 120) != (80 < y * 7 < 120):
        triggered.add(3)
    if (80 < y < 120) != (80 < 100 < 120):
        triggered.add(4)

    if (1.2 < z < 1.5) != (1.2 < z < 15):
        triggered.add(5)
    if (1.2 < z < 1.5) != (1.2 < z < 5):
        triggered.add(6)

    if (x > 60 and x < 70 and y > 90 and y < 110) != (x > 60 and x < 70 and y > 190 and y < 110):
        triggered.add(7)
    if (x > 60 and x < 70 and y > 90 and y < 110) != (x > 60 and x < 90 and y > 90 and y < 110):
        triggered.add(8)

    if (x > 55 and x < 75 and z > 1.25 and z < 1.45) != (x > 55 and x < 75 and z > 1.25 and z < 14.5):
        triggered.add(9)
    if (x > 55 and x < 75 and z > 1.25 and z < 1.45) != (x > 55 and x < 75 and z > 1.25 and z < 15):
        triggered.add(10)

    if (y > 85 and y < 115 and z > 1.25 and z < 1.45) != (y > 85 and y < 115 and z > 1.25 and z < 14.5):
        triggered.add(11)
    if (y > 85 and y < 115 and z > 1.25 and z < 1.45) != (y > 85 and y < 115 and z > 1.25 and z < 145):
        triggered.add(12)

    if (x > 63 and x < 67) != (x > 63 and x < 167):
        triggered.add(13)
    if (x > 63 and x < 67) != (x > 163 and x < 67):
        triggered.add(14)

    if (y > 98 and y < 102) != (y > 98 and y < 12.2):
        triggered.add(15)
    if (y > 98 and y < 102) != (y > 918 and y < 102):
        triggered.add(16)

    if (z > 1.33 and z < 1.37) != (z > 1.33 and z < 13.7):
        triggered.add(17)
    if (z > 1.33 and z < 1.37) != (z > 1.33 and 1 < 1.37):
        triggered.add(18)

    if (abs(x - 65) < 2.5) != (abs(x - 65) < 12.5):
        triggered.add(19)
    if (abs(x - 65) < 2.5) != (abs(x - 65) < 25):
        triggered.add(20)

    if (abs(y - 100) < 5) != (abs(y - 100) < 15):
        triggered.add(21)
    if (abs(y - 100) < 5) != (abs(y - 100) < 51):
        triggered.add(22)

    if (abs(z - 1.35) < 0.05) != (abs(z - 1.35) < 5):
        triggered.add(23)
    if (abs(z - 1.35) < 0.05) != (abs(z - 1.35) < 2.05):
        triggered.add(24)

    if (x > 55 and y > 85 and z > 1.25) != (x > 55 and y > 85 and z > 125):
        triggered.add(25)
    if (x > 55 and y > 85 and z > 1.25) != (x > 55 and y > 85 and 3 > 1.25):
        triggered.add(26)

    if (x < 75 and y < 115 and z < 1.45) != (x < 75 and y < 115 and z < 145):
        triggered.add(27)
    if (x < 75 and y < 115 and z < 1.45) != (x < 75 and y < 115 and z < 14.5):
        triggered.add(28)

    if ((x + y + z * 100) > 280 and (x + y + z * 100) < 320) != (
        (x + y + z * 100) > 480 and (x + y + z * 100) < 320):
        triggered.add(29)
    if ((x + y + z * 100) > 280 and (x + y + z * 100) < 320) != (
        (x + y + z * 100) > 280 and (x + y + z * 100) < 520):
        triggered.add(30)

    if (x * y > 5500 and x * y < 7500) != (x * y > 2500 and x * y < 7500):
        triggered.add(31)
    if (x * y > 5500 and x * y < 7500) != (x * y > 5500 and x * y < 4500):
        triggered.add(32)

    if (x * z > 70 and x * z < 110) != (x * z > 70 and x * z < 920):
        triggered.add(33)
    if (x * z > 70 and x * z < 110) != (x * z > 70 and x * z < 1110):
        triggered.add(34)

    if (y * z > 110 and y * z < 150) != (y * z > 1410 and y * z < 150):
        triggered.add(35)
    if (y * z > 110 and y * z < 150) != (y * z > 110 and y * z < 1550):
        triggered.add(36)

    if ((x + y + z * 100) / 3 > 93 and (x + y + z * 100) / 3 < 107) != (
        (x + y + z * 100) / 3 > 931 and (x + y + z * 100) / 3 < 107):
        triggered.add(37)
    if ((x + y + z * 100) / 3 > 93 and (x + y + z * 100) / 3 < 107) != (
        (x + y + z * 100) / 3 > 93 and (x + y + z * 200) / 3 < 107):
        triggered.add(38)

    if (math.sqrt(x ** 2 + y ** 2 + (z * 100) ** 2) > 150) != (math.sqrt(x ** 2 + y ** 2 + (z * 100) ** 2) > 1510):
        triggered.add(39)
    if (math.sqrt(x ** 2 + y ** 2 + (z * 100) ** 2) > 150) != (math.sqrt(x ** 2 + y ** 2 + (z * 100) ** 2) > 120):
        triggered.add(40)

    if (x / y > 0.55 and x / y < 0.85) != (x / y > 0.55 and x / y < 85):
        triggered.add(41)
    if (x / y > 0.55 and x / y < 0.85) != (x / y > 0.55 and x / y < 8.5):
        triggered.add(42)

    if (x / z > 40 and x / z < 60) != (x / z > 40 and x / z < 610):
        triggered.add(43)
    if (x / z > 40 and x / z < 60) != (x / z > 401 and x / z < 60):
        triggered.add(44)

    if (y / z > 60 and y / z < 90) != (y / z > 60 and y / z < 910):
        triggered.add(45)
    if (y / z > 60 and y / z < 90) != (y / z > 160 and y / z < 90):
        triggered.add(46)

    if ((x - 50) + (y - 80) > 20 and (x - 50) + (y - 80) < 40) != (
        (x - 50) + (y - 80) > 20 and (x - 50) + (y - 80) < 410):
        triggered.add(47)
    if ((x - 50) + (y - 80) > 20 and (x - 50) + (y - 80) < 40) != (
        (x - 50) + (y - 80) > 20 and (x - 50) + (y - 80) < 140):
        triggered.add(48)

    if ((x - 50) + (z - 1.2) * 100 > 25 and (x - 50) + (z - 1.2) * 100 < 45) != (
        (x - 50) + (z - 1.2) * 100 > 25 and (x - 50) + (z - 1.2) * 100 < 145):
        triggered.add(49)
    if ((x - 50) + (z - 1.2) * 100 > 25 and (x - 50) + (z - 1.2) * 100 < 45) != (
        (x - 50) + (z - 1.2) * 100 > 25 and (x - 50) + (z - 1.2) * 100 < 451):
        triggered.add(50)

    if ((y - 80) + (z - 1.2) * 100 > 35 and (y - 80) + (z - 1.2) * 100 < 55) != (
        (y - 80) + (z - 1.2) * 100 > 35 and (y - 80) + (z - 1.2) * 100 < 155):
        triggered.add(51)
    if ((y - 80) + (z - 1.2) * 100 > 35 and (y - 80) + (z - 1.2) * 100 < 55) != (
        (y - 80) + (z - 1.2) * 100 > 35 and (y - 80) + (z - 1.2) * 100 < 515):
        triggered.add(52)

    if (abs((x - 65) - (y - 100) * 0.65) < 8) != (abs((x - 65) - (y - 100) * 0.65) < 18):
        triggered.add(53)
    if (abs((x - 65) - (y - 100) * 0.65) < 8) != (abs((x - 65) - (y - 100) * 0.65) < 81):
        triggered.add(54)

    if (abs((x - 65) - (z - 1.35) * 50) < 10) != (abs((x - 65) - (z - 1.35) * 50) < 110):
        triggered.add(55)
    if (abs((x - 65) - (z - 1.35) * 50) < 10) != (abs((x - 65) - (z - 1.35) * 50) < 101):
        triggered.add(56)

    if (abs((y - 100) - (z - 1.35) * 100) < 12) != (abs((y - 100) - (z - 1.35) * 100) < 112):
        triggered.add(57)
    if (abs((y - 100) - (z - 1.35) * 100) < 12) != (abs((y - 100) - (z - 1.35) * 100) < 132):
        triggered.add(58)

    if (x / (y + 20) > 0.5 and x / (y + 20) < 0.8) != (x / (y + 20) > 0.5 and x / (y + 20) < 8):
        triggered.add(59)
    if (x / (y + 20) > 0.5 and x / (y + 20) < 0.8) != (x / (y + 20) > 0.5 and x / (y + 20) < 48):
        triggered.add(60)

    if (z / (x / 50) > 0.9 and z / (x / 50) < 1.4) != (z / (x / 50) > 0.9 and z / (x / 50) < 14.4):
        triggered.add(61)
    if (z / (x / 50) > 0.9 and z / (x / 50) < 1.4) != (z / (x / 50) > 0.9 and z / (x / 50) < 133.4):
        triggered.add(62)

    if (x * y * z > 7000 and x * y * z < 11000) != (x * y * z > 7000 and x * y * 2 < 11000):
        triggered.add(63)
    if (x * y * z > 7000 and x * y * z < 11000) != (x * y * z > 7000 and x * y * 54 < 11000):
        triggered.add(64)

    if (x * 0.6 + y * 0.3 + z * 10 > 65 and x * 0.6 + y * 0.3 + z * 10 < 75) != (
        x * 0.6 + y * 0.3 + z * 10 > 645 and x * 0.6 + y * 0.3 + z * 10 < 75):
        triggered.add(65)
    if (x * 0.6 + y * 0.3 + z * 10 > 65 and x * 0.6 + y * 0.3 + z * 10 < 75) != (
        x * 0.6 + y * 40.3 + z * 10 > 65 and x * 0.6 + y * 0.3 + z * 10 < 75):
        triggered.add(66)

    if ((x / 65) ** 0.5 * (y / 100) ** 0.3 * (z / 1.35) ** 0.2 > 0.92) != (
        (x / 65) ** 0.5 * (y / 100) ** 0.3 * (z / 1.35) ** 0.2 > 392):
        triggered.add(67)
    if ((x / 65) ** 0.5 * (y / 100) ** 0.3 * (z / 1.35) ** 0.2 > 0.92) != (
        (x / 65) ** 0.5 * (y / 100) ** 0.3 * (z / 1.35) ** 0.2 > 982):
        triggered.add(68)

    if ((x - 65) * (y - 100) > -150 and (x - 65) * (y - 100) < 150) != (
        (x - 65) * (y - 100) > -150 and (x - 65) * (y - 100) < 1250):
        triggered.add(69)
    if ((x - 65) * (y - 100) > -150 and (x - 65) * (y - 100) < 150) != (
        (x - 65) * (y - 100) > -150 and (x - 65) * (y - 100) < 1050):
        triggered.add(70)

    if ((x - 65) * (z - 1.35) > -8 and (x - 65) * (z - 1.35) < 8) != (
        (x - 65) * (z - 1.35) > -8 and (x - 65) * (z - 1.35) < 18):
        triggered.add(71)
    if ((x - 65) * (z - 1.35) > -8 and (x - 65) * (z - 1.35) < 8) != (
        (x - 65) * (z - 1.35) > -8 and (x - 65) * (z - 1.35) < 448):
        triggered.add(72)

    if (abs(x / 65 + y / 100 + z / 1.35 - 3) < 0.25) != (abs(x / 65 + y / 100 + z / 1.35 - 3) < 25):
        triggered.add(73)
    if (abs(x / 65 + y / 100 + z / 1.35 - 3) < 0.25) != (abs(x / 65 + y / 100 + z / 1.35 - 3) < 40.25):
        triggered.add(74)

    if (abs((x / 65) * (y / 100) * (z / 1.35) - 1) < 0.18) != (abs((x / 65) * (y / 100) * (z / 1.35) - 1) < 40.18):
        triggered.add(75)
    if (abs((x / 65) * (y / 100) * (z / 1.35) - 1) < 0.18) != (abs((x / 65) * (y / 100) * (z / 1.35) - 1) < 18):
        triggered.add(76)

    if (x > 62 and x < 68 and y > 97 and y < 103 and z > 1.32 and z < 1.38) != (
        x > 62 and x < 68 and y > 97 and y < 103 and z > 1.32 and z < 138):
        triggered.add(77)
    if (x > 62 and x < 68 and y > 97 and y < 103 and z > 1.32 and z < 1.38) != (
        x > 62 and x < 68 and y > 97 and y < 103 and z > 1.32 and z < 228):
        triggered.add(78)

    if (abs(x / y - 0.65) < 0.08 and abs(z - 1.35) < 0.03) != (abs(x / y - 0.65) < 0.08 and abs(z - 1.35) < 3):
        triggered.add(79)
    if (abs(x / y - 0.65) < 0.08 and abs(z - 1.35) < 0.03) != (abs(x / y - 0.65) < 0.08 and abs(z - 1.35) < 40.03):
        triggered.add(80)

    if (x * y > 6300 and x * y < 6700 and z > 1.33 and z < 1.37) != (
        x * y > 6300 and x * y < 6700 and z > 1.33 and z < 41.37):
        triggered.add(81)
    if (x * y > 6300 and x * y < 6700 and z > 1.33 and z < 1.37) != (
        x * y > 6300 and x * y < 6700 and z > 1.33 and z < 91.37):
        triggered.add(82)

    if ((x + y) / 2 > 80 and (x + y) / 2 < 86 and z > 1.32 and z < 1.38) != (
        (x + y) / 2 > 80 and (x + y) / 2 < 86 and z > 1.32 and z < 138):
        triggered.add(83)
    if ((x + y) / 2 > 80 and (x + y) / 2 < 86 and z > 1.32 and z < 1.38) != (
        (x + y) / 2 > 80 and (x + y) / 2 < 86 and z > 1.32 and z < 19.38):
        triggered.add(84)

    if (abs(x - y) < 40 and z > 1.32 and z < 1.38) != (abs(x - y) < 40 and z > 1.32 and z < 138):
        triggered.add(85)
    if (abs(x - y) < 40 and z > 1.32 and z < 1.38) != (abs(x - y) < 40 and z > 1.32 and z < 13.8):
        triggered.add(86)

    if (math.sqrt((x - 65) ** 2 + (y - 100) ** 2) < 5 and abs(z - 1.35) < 0.03) != (
        math.sqrt((x - 65) ** 2 + (y - 100) ** 2) < 5 and abs(z - 1.35) < 13):
        triggered.add(87)
    if (math.sqrt((x - 65) ** 2 + (y - 100) ** 2) < 5 and abs(z - 1.35) < 0.03) != (
        math.sqrt((x - 65) ** 2 + (y - 100) ** 2) < 5 and abs(z - 1.35) < 8):
        triggered.add(88)

    if (x / 65 > 0.975 and x / 65 < 1.025 and y / 100 > 0.97 and y / 100 < 1.03) != (
        x / 65 > 0.975 and x / 65 < 1.025 and y / 100 > 10.97 and y / 100 < 1.03):
        triggered.add(89)
    if (x / 65 > 0.975 and x / 65 < 1.025 and y / 100 > 0.97 and y / 100 < 1.03) != (
        x / 65 > 0.975 and x / 65 < 1.025 and y / 100 > 40.97 and y / 100 < 1.03):
        triggered.add(90)

    if (z / 1.35 > 0.978 and z / 1.35 < 1.022) != (z / 1.35 > 0.978 and z / 1.35 < 41.022):
        triggered.add(91)
    if (z / 1.35 > 0.978 and z / 1.35 < 1.022) != (z / 1.35 > 0.978 and z / 1.35 < 622):
        triggered.add(92)

    if ((x / 65 + y / 100 + z / 1.35) / 3 > 0.975 and (x / 65 + y / 100 + z / 1.35) / 3 < 1.025) != (
        (x / 65 + y / 100 + z / 1.35) / 3 > 9.75 and (x / 65 + y / 100 + z / 1.35) / 3 < 1.025):
        triggered.add(93)
    if ((x / 65 + y / 100 + z / 1.35) / 3 > 0.975 and (x / 65 + y / 100 + z / 1.35) / 3 < 1.025) != (
        (x / 65 + y / 100 + z / 1.35) / 3 > 0.975 and (x / 65 + y / 100 + z / 1.35) / 3 < 10.25):
        triggered.add(94)

    if (max(abs(x / 65 - 1), abs(y / 100 - 1), abs(z / 1.35 - 1)) < 0.025) != (
        max(abs(x / 65 - 1), abs(y / 100 - 1), abs(z / 1.35 - 1)) < 40.025):
        triggered.add(95)
    if (max(abs(x / 65 - 1), abs(y / 100 - 1), abs(z / 1.35 - 1)) < 0.025) != (
        max(abs(x / 65 - 1), abs(y / 100 - 1), abs(z / 1.35 - 1)) < 10.025):
        triggered.add(96)

    if (min(x / 65, y / 100, z / 1.35) > 0.975) != (min(x / 65, y / 100, z / 1.35) > 9.75):
        triggered.add(97)
    if (min(x / 65, y / 100, z / 1.35) > 0.975) != (min(x / 65, y / 100, z / 1.35) > 19.75):
        triggered.add(98)

    if (max(x / 65, y / 100, z / 1.35) < 1.025) != (max(x / 65, y / 100, z / 1.35) < 11.025):
        triggered.add(99)
    if (max(x / 65, y / 100, z / 1.35) < 1.025) != (max(x / 65, y / 100, z / 1.35) < 41.025):
        triggered.add(100)

    if (abs(max(x, y, z * 100) - min(x, y, z * 100)) < 40) != (abs(max(x, y, z * 100) - min(x, y, z * 100)) < 140):
        triggered.add(101)
    if (abs(max(x, y, z * 100) - min(x, y, z * 100)) < 40) != (abs(max(x, y, z * 100) - min(x, y, z * 100)) < 1140):
        triggered.add(102)

    if ((x + y + z * 100) > 298 and (x + y + z * 100) < 302) != (
        (x + y + z * 100) > 298 and (x + y + z * 100) < 3102):
        triggered.add(103)
    if ((x + y + z * 100) > 298 and (x + y + z * 100) < 302) != (
        (x + y + z * 100) > 298 and (x + y + z * 100) < 352):
        triggered.add(104)

    if (x * y * z > 8700 and x * y * z < 9300) != (x * y * z > 8700 and x * y * 4 < 9300):
        triggered.add(105)
    if (x * y * z > 8700 and x * y * z < 9300) != (x * y * z > 8700 and x * y * z < 5300):
        triggered.add(106)

    if (abs((x + y + z * 100) / 3 - 100) < 2) != (abs((x + y + z * 100) / 3 - 100) < 332):
        triggered.add(107)
    if (abs((x + y + z * 100) / 3 - 100) < 2) != (abs((x + y + z * 100) / 3 - 100) < 122):
        triggered.add(108)

    if (z > 1.345 and z < 1.355) != (z > 1.345 and z < 1355):
        triggered.add(109)
    if (z > 1.345 and z < 1.355) != (z > 1.345 and z < 355):
        triggered.add(110)

    if (x < 52 or y < 85 or z < 1.22) != (x < 52 or y < 85 or z < 122):
        triggered.add(111)
    if (x < 52 or y < 85 or z < 1.22) != (x < 52 or y < 855 or z < 1.22):
        triggered.add(112)

    if (x > 78 or y > 115 or z > 1.48) != (x > 78 or y > 115 or z > 148):
        triggered.add(113)
    if (x > 78 or y > 115 or z > 1.48) != (x > 78 or y > 115 or z * 8 > 1.48):
        triggered.add(114)

    if (abs(x / y - 0.65) > 0.12) != (abs(x / y - 0.65) > 12):
        triggered.add(115)
    if (abs(x / y - 0.65) > 0.12) != (abs(x / y - 0.65) > 20.12):
        triggered.add(116)

    if (abs(x / z - 48) > 6) != (abs(x / z - 48) > 16):
        triggered.add(117)
    if (abs(x / z - 48) > 6) != (abs(x / z - 48) > 36):
        triggered.add(118)

    if (abs(y / z - 74) > 8) != (abs(y / z - 74) > 228):
        triggered.add(119)
    if (abs(y / z - 74) > 8) != (abs(y / z - 74) > 82):
        triggered.add(120)

    if ((x + y + z * 100) < 285 or (x + y + z * 100) > 315) != (
        (x + y + z * 100) < 285 or (x + y + z * 100) > 3165):
        triggered.add(121)
    if ((x + y + z * 100) < 285 or (x + y + z * 100) > 315) != (
        (x + y + z * 100) < 2835 or (x + y + z * 100) > 315):
        triggered.add(122)

    if (x * y * z < 8000 or x * y * z > 10000) != (x * y * z < 8000 or x * y * z > 1000):
        triggered.add(123)
    if (x * y * z < 8000 or x * y * z > 10000) != (x * y * z < 8000 or x * y * z > 93000):
        triggered.add(124)

    if (x < 55 and y < 90) != (x < 55 and y < 290):
        triggered.add(125)
    if (x < 55 and y < 90) != (x < 55 and 80 < 90):
        triggered.add(126)

    if (x > 75 and y > 110) != (x > 75 and 200 > 110):
        triggered.add(127)
    if (x > 75 and y > 110) != (x > 75 and 500 > 110):
        triggered.add(128)

    if (x < 55 and z < 1.25) != (x < 55 and z < 125):
        triggered.add(129)
    if (x < 55 and z < 1.25) != (x < 55 and z < 12.5):
        triggered.add(130)

    if (x > 75 and z > 1.45) != (x > 75 and z > 145):
        triggered.add(131)
    if (x > 75 and z > 1.45) != (x > 75 and 6 > 1.45):
        triggered.add(132)

    if (y < 90 and z < 1.25) != (y < 90 and 1 < 1.25):
        triggered.add(133)
    if (y < 90 and z < 1.25) != (y < 90 and z < 125):
        triggered.add(134)

    if (y < 75 or y > 125) != (y < 75 or y * 8 > 125):
        triggered.add(135)
    if (y < 75 or y > 125) != (y < 75 or y * 10 > 125):
        triggered.add(136)

    if (x < 45 and y < 80 and z < 1.2) != (x < 45 and y < 80 and z < 12):
        triggered.add(137)
    if (x < 45 and y < 80 and z < 1.2) != (x < 45 and y < 80 and z < 11.2):
        triggered.add(138)

    if (abs((x * y * z) / 9000 - 1) > 0.22) != (abs((x * y * z) / 9000 - 1) > 9.22):
        triggered.add(139)
    if (abs((x * y * z) / 9000 - 1) > 0.22) != (abs((x * y * z) / 9000 - 1) > 22):
        triggered.add(140)

    return triggered

def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """"""
    generated_path = execute_validation_rules(particle[0], particle[1], particle[2])

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # : x:1-50, y:1-50, z:1-50
        self.bounds = bounds if bounds else [(1, 80), (1, 120), (1, 15)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """target pathsPSO"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(particles[i][0], particles[i][1], particles[i][2]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(gbest_particle[0], gbest_particle[1], gbest_particle[2]),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """Path PSO"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - Path ")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations")
    print(f"Path : {len(target_paths)}")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"Path {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "" if result['success'] else f"({result['best_fitness']:.3f})"
        print(f"{status} | {result['time']:.2f}s | iterations{result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f": {success_count}/{len(target_paths)} ({success_rate:.1f}%) | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs}")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations, {len(target_paths)}Path ")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"---  {run_idx}/{num_runs} ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f": {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs} runcompleted | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 1: 
    ws1 = wb.active
    ws1.title = ""
    ws1.sheet_view.showGridLines = False

    headers = ["Run", "", "", "", "Average Iterations", "(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f" {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # 2: Path 
    ws2 = wb.create_sheet(title="Path ")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Path ID", "", "", "", "Average Iterations", "Minimum Iterations", "Maximum Iterations"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"Path  {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # 3: 
    ws3 = wb.create_sheet(title="")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Path ", "", "(x,y,z)", "", "Iterations", "Path "]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"{run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # 4: target paths
    ws4 = wb.create_sheet(title="target paths")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Path ID", "target paths", ""]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"Path  {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f" : {filename}")
    print(f"{'=' * 70}")
    print(f":")
    print(f"  1.        - {len(all_results)} run")
    print(f"  2. Path        - Path ")
    print(f"  3.    -  runPath ")
    print(f"  4. target paths       - target paths")
    print(f"{'=' * 70}\n")

    return filename


def main():


    target_paths = [
        {1, 3, 5, 6, 7, 9, 10, 11, 12, 13, 17, 18, 19, 20, 23, 24, 25, 27, 28, 30, 32, 33, 34, 36, 39, 47, 48, 49, 50,
         51,
         52, 55, 56, 57, 58, 61, 62, 67, 68, 73, 74, 75, 76, 79, 80, 83, 84, 85, 86, 87, 88, 91, 92, 94, 95, 96, 97, 98,
         99,
         100, 101, 102, 103, 107, 108, 109, 110, 111, 112, 113, 117, 118, 119, 120, 121, 124, 135, 136, 139, 140},

        {1, 3, 5, 6, 7, 9, 10, 11, 12, 13, 15, 16, 17, 18, 19, 20, 23, 24, 25, 27, 28, 30, 32, 33, 34, 36, 39, 49, 50,
         51,
         52, 55, 56, 57, 58, 61, 62, 67, 68, 73, 74, 75, 76, 79, 80, 83, 84, 85, 86, 87, 88, 91, 92, 94, 95, 96, 97, 98,
         99,
         100, 101, 102, 103, 107, 108, 109, 110, 111, 112, 113, 117, 118, 119, 120, 121, 124, 135, 136, 139, 140},

        {1, 3, 5, 6, 9, 10, 11, 12, 13, 17, 18, 19, 20, 21, 22, 23, 24, 25, 27, 28, 30, 32, 33, 34, 36, 39, 49, 50, 51,
         52,
         53, 54, 55, 56, 57, 58, 61, 62, 67, 68, 73, 74, 75, 76, 81, 82, 83, 84, 85, 86, 91, 92, 94, 95, 96, 99, 100,
         101,
         102, 103, 107, 108, 109, 110, 111, 112, 113, 115, 116, 117, 118, 119, 120, 121, 124, 135, 136, 139, 140},

        {2, 3, 5, 6, 11, 12, 13, 17, 18, 20, 21, 22, 23, 24, 25, 30, 32, 33, 34, 36, 39, 41, 42, 49, 50, 51, 52, 54, 55,
         56,
         57, 58, 67, 68, 71, 72, 73, 74, 75, 76, 83, 84, 85, 86, 91, 92, 94, 95, 96, 99, 100, 101, 102, 103, 107, 108,
         109,
         110, 111, 112, 115, 116, 117, 118, 119, 120, 121, 124, 127, 128, 131, 133, 134, 135, 136, 139, 140},

        {1, 3, 5, 6, 8, 9, 10, 11, 12, 13, 17, 18, 19, 20, 21, 22, 23, 24, 25, 27, 28, 30, 32, 33, 34, 36, 39, 49, 50,
         51,
         52, 53, 54, 55, 56, 57, 58, 67, 68, 73, 74, 75, 76, 81, 82, 83, 84, 85, 86, 91, 92, 94, 95, 96, 99, 100, 101,
         102,
         103, 107, 108, 109, 110, 111, 112, 113, 115, 116, 117, 118, 119, 120, 121, 124, 135, 136, 139, 140},

        {1, 3, 5, 6, 7, 9, 10, 11, 12, 15, 16, 17, 18, 23, 24, 25, 27, 28, 30, 32, 33, 34, 36, 39, 49, 50, 51, 52, 55,
         56,
         57, 58, 61, 62, 67, 68, 73, 74, 75, 76, 77, 78, 79, 80, 83, 84, 85, 86, 87, 88, 91, 92, 94, 95, 96, 99, 100,
         101,
         102, 103, 107, 108, 109, 110, 111, 112, 113, 118, 119, 120, 121, 124, 135, 136, 139, 140},

        {1, 3, 5, 6, 9, 10, 11, 12, 14, 17, 18, 21, 22, 23, 24, 25, 27, 28, 30, 32, 33, 34, 36, 39, 47, 48, 49, 50, 51,
         52,
         53, 54, 55, 56, 57, 58, 61, 62, 67, 68, 73, 74, 75, 76, 79, 80, 91, 92, 94, 95, 96, 97, 98, 99, 100, 101, 102,
         103,
         107, 108, 109, 110, 111, 112, 113, 117, 118, 119, 120, 121, 124, 135, 136, 139, 140},

        {1, 3, 5, 6, 9, 10, 13, 17, 18, 19, 20, 22, 23, 24, 27, 28, 30, 32, 33, 34, 36, 39, 41, 42, 49, 50, 51, 52, 53,
         54,
         55, 56, 57, 58, 61, 62, 67, 68, 73, 74, 75, 76, 85, 86, 91, 92, 94, 95, 96, 99, 100, 101, 102, 103, 104, 107,
         108,
         109, 110, 113, 115, 116, 117, 118, 119, 120, 121, 124, 133, 134, 135, 136, 139, 140},

        {1, 4, 5, 6, 13, 17, 18, 20, 22, 23, 24, 30, 32, 33, 34, 36, 39, 41, 42, 49, 50, 51, 52, 54, 55, 56, 57, 58, 59,
         60,
         67, 68, 71, 72, 73, 74, 75, 76, 85, 86, 91, 92, 94, 95, 96, 99, 100, 101, 102, 103, 107, 108, 109, 110, 113,
         115,
         116, 117, 118, 119, 120, 121, 124, 127, 128, 131, 133, 134, 135, 136, 139, 140},

        {1, 3, 5, 6, 11, 12, 13, 17, 18, 20, 21, 22, 23, 24, 25, 30, 33, 34, 36, 39, 47, 48, 49, 50, 51, 52, 55, 56, 57,
         58,
         67, 68, 69, 70, 71, 72, 73, 74, 75, 76, 79, 80, 85, 86, 91, 92, 94, 95, 96, 97, 98, 99, 100, 101, 102, 103,
         107,
         108, 109, 110, 111, 112, 117, 118, 119, 120, 121, 124, 131, 135, 136, 139, 140},

        {1, 4, 5, 6, 9, 10, 13, 17, 18, 19, 20, 22, 23, 24, 27, 28, 30, 31, 33, 34, 36, 39, 41, 42, 49, 50, 51, 52, 54,
         55,
         56, 57, 58, 61, 62, 64, 67, 68, 73, 74, 75, 76, 85, 86, 91, 92, 94, 95, 96, 99, 100, 101, 102, 103, 104, 107,
         108,
         109, 110, 113, 115, 116, 117, 118, 119, 120, 121, 124, 133, 134, 135, 136},
        {1, 3, 5, 7, 9, 10, 11, 12, 14, 15, 16, 17, 18, 23, 25, 27, 28, 32, 33, 34, 36, 39, 50, 52, 61, 62, 67, 68, 73,
         74,
         75, 76, 77, 78, 80, 81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 94, 95, 96, 97, 98, 99, 100, 102, 103, 107,
         108, 109, 110, 111, 112, 113, 118, 119, 120, 121, 124, 135, 136, 139, 140},

        {1, 4, 5, 6, 13, 17, 18, 19, 20, 22, 23, 24, 30, 32, 33, 34, 35, 39, 41, 42, 49, 50, 51, 52, 54, 55, 56, 57, 58,
         59,
         60, 67, 68, 73, 74, 75, 76, 85, 86, 91, 92, 94, 95, 96, 99, 100, 101, 102, 103, 104, 107, 108, 109, 110, 113,
         115,
         116, 117, 118, 119, 120, 121, 124, 127, 128, 131, 133, 134, 139, 140},
        {1, 3, 5, 6, 11, 12, 17, 18, 20, 21, 22, 23, 24, 27, 28, 30, 32, 36, 39, 49, 50, 51, 52, 54, 55, 56, 57, 58, 61,
         62,
         67, 68, 73, 74, 75, 76, 83, 84, 91, 92, 94, 95, 96, 99, 100, 102, 103, 107, 108, 109, 110, 111, 112, 113, 115,
         116,
         118, 119, 120, 121, 124, 125, 126, 129, 130, 135, 136, 139, 140},

        {1, 3, 5, 6, 11, 12, 17, 18, 19, 20, 21, 22, 23, 24, 27, 28, 30, 31, 36, 39, 49, 50, 51, 52, 55, 56, 57, 58, 61,
         62,
         64, 67, 68, 69, 70, 79, 80, 85, 86, 91, 92, 94, 95, 96, 99, 100, 102, 103, 104, 105, 106, 107, 108, 109, 110,
         111,
         112, 113, 118, 119, 120, 121, 123, 129, 130, 133, 134, 135, 136},

        {1, 4, 5, 6, 9, 10, 13, 17, 18, 19, 20, 22, 23, 24, 27, 28, 30, 31, 33, 34, 36, 39, 41, 42, 50, 51, 52, 54, 55,
         56,
         59, 60, 61, 62, 63, 67, 68, 71, 72, 73, 74, 75, 76, 85, 86, 91, 92, 94, 95, 96, 99, 100, 102, 103, 107, 108,
         109,
         110, 113, 115, 116, 118, 119, 120, 121, 124, 133, 134, 139, 140},

        {1, 4, 5, 6, 17, 18, 19, 20, 22, 23, 24, 27, 28, 29, 31, 35, 37, 38, 39, 41, 42, 49, 50, 51, 52, 53, 54, 55, 56,
         57,
         58, 61, 62, 65, 69, 70, 75, 76, 85, 86, 91, 92, 93, 95, 96, 99, 100, 102, 103, 104, 107, 108, 109, 110, 113,
         115,
         116, 118, 119, 120, 121, 129, 130, 133, 134, 139, 140},

        {2, 4, 5, 6, 17, 18, 22, 23, 24, 27, 28, 29, 31, 36, 37, 38, 39, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 61, 62,
         66,
         69, 70, 75, 76, 85, 86, 91, 92, 95, 96, 99, 100, 102, 103, 104, 107, 108, 109, 110, 113, 115, 116, 118, 119,
         120,
         122, 129, 130, 133, 134, 135, 136, 137, 138, 139, 140},

        {1, 4, 13, 19, 20, 22, 23, 24, 26, 29, 35, 37, 38, 39, 43, 45, 47, 48, 55, 56, 57, 58, 63, 64, 67, 68, 69, 70,
         79,
         80, 94, 95, 96, 99, 100, 101, 102, 105, 106, 107, 108, 118, 119, 120, 122, 123, 135, 136},

        {1, 3, 13, 20, 21, 22, 23, 24, 26, 29, 35, 37, 38, 39, 43, 45, 47, 48, 55, 56, 57, 58, 63, 64, 67, 68, 69, 70,
         79,
         80, 93, 95, 96, 99, 100, 105, 106, 107, 108, 118, 119, 120, 122, 123, 132, 135, 136},

        {1, 3, 8, 13, 19, 20, 21, 22, 23, 24, 26, 29, 37, 38, 39, 43, 45, 47, 48, 55, 56, 57, 58, 63, 64, 67, 68, 79,
         80,
         93, 95, 96, 99, 100, 107, 108, 114, 118, 119, 120, 123, 127, 128, 132, 135, 136},

        {2, 3, 21, 22, 23, 24, 31, 40, 45, 54, 55, 56, 57, 58, 61, 62, 66, 71, 72, 73, 74, 75, 76, 95, 96, 99, 100, 101,
         102, 107, 108, 114, 115, 116, 117, 118, 119, 120, 125, 126, 135, 136, 139, 140},

        {2, 3, 20, 21, 22, 23, 24, 31, 40, 44, 45, 53, 54, 57, 58, 66, 69, 70, 71, 72, 73, 74, 75, 76, 95, 96, 101, 102,
         107, 108, 114, 115, 116, 117, 118, 119, 120, 125, 126, 135, 136, 139, 140},

        {2, 3, 21, 22, 23, 24, 31, 40, 46, 54, 55, 56, 57, 58, 61, 62, 66, 69, 70, 71, 72, 73, 74, 75, 76, 95, 96, 101,
         102,
         107, 108, 114, 115, 116, 117, 118, 119, 120, 135, 136, 139, 140},
    ]

    print("=" * 70)
    print("baseline PSO")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("Program completed")


if __name__ == "__main__":
    main()