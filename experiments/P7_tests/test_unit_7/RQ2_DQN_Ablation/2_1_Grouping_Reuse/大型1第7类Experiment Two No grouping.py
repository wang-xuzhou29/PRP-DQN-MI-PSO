import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import math

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === 全局取值范围（修改为 X:1~80, Y:1~120, Z:1~15） ===
MIN_X = 1
MAX_X = 80
MIN_Y = 1
MAX_Y = 120
MIN_Z = 1
MAX_Z = 15

# === 归一化/反归一化 ===
def normalize_state(state):
    """将状态归一化到 [0, 1] 区间"""
    weather_norm = (state[0] - MIN_X) / (MAX_X - MIN_X)
    time_norm = (state[1] - MIN_Y) / (MAX_Y - MIN_Y)
    z_norm = (state[2] - MIN_Z) / (MAX_Z - MIN_Z)
    return (weather_norm, time_norm, z_norm)

def denormalize_state(state_norm):
    """将归一化状态还原"""
    weather = int(round(state_norm[0] * (MAX_X - MIN_X) + MIN_X))
    time_period = int(round(state_norm[1] * (MAX_Y - MIN_Y) + MIN_Y))
    z = int(round(state_norm[2] * (MAX_Z - MIN_Z) + MIN_Z))

    # 边界保护
    weather = np.clip(weather, MIN_X, MAX_X)
    time_period = np.clip(time_period, MIN_Y, MAX_Y)
    z = np.clip(z, MIN_Z, MAX_Z)

    return (weather, time_period, z)

def normalize_value(value, min_val, max_val):
    return (value - min_val) / (max_val - min_val)

def denormalize_value(value_norm, min_val, max_val):
    return int(round(value_norm * (max_val - min_val) + min_val))

# === 安全除法 ===
def safe_divide(numerator, denominator, default=0.0):
    if denominator == 0:
        return default
    return numerator / denominator

# === reward function ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10

    if target_path.issubset(triggered):
        reward += 1

    if prev_triggered is not None:
        prev_sim = jaccard_similarity(prev_triggered, target_path)
        improvement = sim - prev_sim
        reward += improvement * 5

    return reward

# ========== 规则触发函数（section7_flow_pressure_density_hybrid） ==========
def section7_flow_pressure_density_hybrid(x, y, z):
    actions = []
    triggered = set()

    if (50 < x < 80) != (50 < x * 8 < 80):
        triggered.add(1)
    if (50 < x < 80) != (50 < 70 < 80):
        triggered.add(2)

    if (80 < y < 120) != (80 < y * 7 < 120):
        triggered.add(3)
    if (80 < y < 120) != (80 < 100 < 120):
        triggered.add(4)

    if (1.2 < z < 1.5) != (1.2 < z < 15):
        triggered.add(5)
    if (1.2 < z < 1.5) != (1.2 < z < 5):
        triggered.add(6)

    if (x > 60 and x < 70 and y > 90 and y < 110) != (x > 60 and x < 70 and y > 190 and y < 110):
        triggered.add(7)
    if (x > 60 and x < 70 and y > 90 and y < 110) != (x > 60 and x < 90 and y > 90 and y < 110):
        triggered.add(8)

    if (x > 55 and x < 75 and z > 1.25 and z < 1.45) != (x > 55 and x < 75 and z > 1.25 and z < 14.5):
        triggered.add(9)
    if (x > 55 and x < 75 and z > 1.25 and z < 1.45) != (x > 55 and x < 75 and z > 1.25 and z < 15):
        triggered.add(10)

    if (y > 85 and y < 115 and z > 1.25 and z < 1.45) != (y > 85 and y < 115 and z > 1.25 and z < 14.5):
        triggered.add(11)
    if (y > 85 and y < 115 and z > 1.25 and z < 1.45) != (y > 85 and y < 115 and z > 1.25 and z < 145):
        triggered.add(12)

    if (x > 63 and x < 67) != (x > 63 and x < 167):
        triggered.add(13)
    if (x > 63 and x < 67) != (x > 163 and x < 67):
        triggered.add(14)

    if (y > 98 and y < 102) != (y > 98 and y < 12.2):
        triggered.add(15)
    if (y > 98 and y < 102) != (y > 918 and y < 102):
        triggered.add(16)

    if (z > 1.33 and z < 1.37) != (z > 1.33 and z < 13.7):
        triggered.add(17)
    if (z > 1.33 and z < 1.37) != (z > 1.33 and 1 < 1.37):
        triggered.add(18)

    if (abs(x - 65) < 2.5) != (abs(x - 65) < 12.5):
        triggered.add(19)
    if (abs(x - 65) < 2.5) != (abs(x - 65) < 25):
        triggered.add(20)

    if (abs(y - 100) < 5) != (abs(y - 100) < 15):
        triggered.add(21)
    if (abs(y - 100) < 5) != (abs(y - 100) < 51):
        triggered.add(22)

    if (abs(z - 1.35) < 0.05) != (abs(z - 1.35) < 5):
        triggered.add(23)
    if (abs(z - 1.35) < 0.05) != (abs(z - 1.35) < 2.05):
        triggered.add(24)

    if (x > 55 and y > 85 and z > 1.25) != (x > 55 and y > 85 and z > 125):
        triggered.add(25)
    if (x > 55 and y > 85 and z > 1.25) != (x > 55 and y > 85 and 3 > 1.25):
        triggered.add(26)

    if (x < 75 and y < 115 and z < 1.45) != (x < 75 and y < 115 and z < 145):
        triggered.add(27)
    if (x < 75 and y < 115 and z < 1.45) != (x < 75 and y < 115 and z < 14.5):
        triggered.add(28)

    if ((x + y + z * 100) > 280 and (x + y + z * 100) < 320) != (
        (x + y + z * 100) > 480 and (x + y + z * 100) < 320):
        triggered.add(29)
    if ((x + y + z * 100) > 280 and (x + y + z * 100) < 320) != (
        (x + y + z * 100) > 280 and (x + y + z * 100) < 520):
        triggered.add(30)

    if (x * y > 5500 and x * y < 7500) != (x * y > 2500 and x * y < 7500):
        triggered.add(31)
    if (x * y > 5500 and x * y < 7500) != (x * y > 5500 and x * y < 4500):
        triggered.add(32)

    if (x * z > 70 and x * z < 110) != (x * z > 70 and x * z < 920):
        triggered.add(33)
    if (x * z > 70 and x * z < 110) != (x * z > 70 and x * z < 1110):
        triggered.add(34)

    if (y * z > 110 and y * z < 150) != (y * z > 1410 and y * z < 150):
        triggered.add(35)
    if (y * z > 110 and y * z < 150) != (y * z > 110 and y * z < 1550):
        triggered.add(36)

    if ((x + y + z * 100) / 3 > 93 and (x + y + z * 100) / 3 < 107) != (
        (x + y + z * 100) / 3 > 931 and (x + y + z * 100) / 3 < 107):
        triggered.add(37)
    if ((x + y + z * 100) / 3 > 93 and (x + y + z * 100) / 3 < 107) != (
        (x + y + z * 100) / 3 > 93 and (x + y + z * 200) / 3 < 107):
        triggered.add(38)

    if (math.sqrt(x ** 2 + y ** 2 + (z * 100) ** 2) > 150) != (math.sqrt(x ** 2 + y ** 2 + (z * 100) ** 2) > 1510):
        triggered.add(39)
    if (math.sqrt(x ** 2 + y ** 2 + (z * 100) ** 2) > 150) != (math.sqrt(x ** 2 + y ** 2 + (z * 100) ** 2) > 120):
        triggered.add(40)

    if (x / y > 0.55 and x / y < 0.85) != (x / y > 0.55 and x / y < 85):
        triggered.add(41)
    if (x / y > 0.55 and x / y < 0.85) != (x / y > 0.55 and x / y < 8.5):
        triggered.add(42)

    if (x / z > 40 and x / z < 60) != (x / z > 40 and x / z < 610):
        triggered.add(43)
    if (x / z > 40 and x / z < 60) != (x / z > 401 and x / z < 60):
        triggered.add(44)

    if (y / z > 60 and y / z < 90) != (y / z > 60 and y / z < 910):
        triggered.add(45)
    if (y / z > 60 and y / z < 90) != (y / z > 160 and y / z < 90):
        triggered.add(46)

    if ((x - 50) + (y - 80) > 20 and (x - 50) + (y - 80) < 40) != (
        (x - 50) + (y - 80) > 20 and (x - 50) + (y - 80) < 410):
        triggered.add(47)
    if ((x - 50) + (y - 80) > 20 and (x - 50) + (y - 80) < 40) != (
        (x - 50) + (y - 80) > 20 and (x - 50) + (y - 80) < 140):
        triggered.add(48)

    if ((x - 50) + (z - 1.2) * 100 > 25 and (x - 50) + (z - 1.2) * 100 < 45) != (
        (x - 50) + (z - 1.2) * 100 > 25 and (x - 50) + (z - 1.2) * 100 < 145):
        triggered.add(49)
    if ((x - 50) + (z - 1.2) * 100 > 25 and (x - 50) + (z - 1.2) * 100 < 45) != (
        (x - 50) + (z - 1.2) * 100 > 25 and (x - 50) + (z - 1.2) * 100 < 451):
        triggered.add(50)

    if ((y - 80) + (z - 1.2) * 100 > 35 and (y - 80) + (z - 1.2) * 100 < 55) != (
        (y - 80) + (z - 1.2) * 100 > 35 and (y - 80) + (z - 1.2) * 100 < 155):
        triggered.add(51)
    if ((y - 80) + (z - 1.2) * 100 > 35 and (y - 80) + (z - 1.2) * 100 < 55) != (
        (y - 80) + (z - 1.2) * 100 > 35 and (y - 80) + (z - 1.2) * 100 < 515):
        triggered.add(52)

    if (abs((x - 65) - (y - 100) * 0.65) < 8) != (abs((x - 65) - (y - 100) * 0.65) < 18):
        triggered.add(53)
    if (abs((x - 65) - (y - 100) * 0.65) < 8) != (abs((x - 65) - (y - 100) * 0.65) < 81):
        triggered.add(54)

    if (abs((x - 65) - (z - 1.35) * 50) < 10) != (abs((x - 65) - (z - 1.35) * 50) < 110):
        triggered.add(55)
    if (abs((x - 65) - (z - 1.35) * 50) < 10) != (abs((x - 65) - (z - 1.35) * 50) < 101):
        triggered.add(56)

    if (abs((y - 100) - (z - 1.35) * 100) < 12) != (abs((y - 100) - (z - 1.35) * 100) < 112):
        triggered.add(57)
    if (abs((y - 100) - (z - 1.35) * 100) < 12) != (abs((y - 100) - (z - 1.35) * 100) < 132):
        triggered.add(58)

    if (x / (y + 20) > 0.5 and x / (y + 20) < 0.8) != (x / (y + 20) > 0.5 and x / (y + 20) < 8):
        triggered.add(59)
    if (x / (y + 20) > 0.5 and x / (y + 20) < 0.8) != (x / (y + 20) > 0.5 and x / (y + 20) < 48):
        triggered.add(60)

    if (z / (x / 50) > 0.9 and z / (x / 50) < 1.4) != (z / (x / 50) > 0.9 and z / (x / 50) < 14.4):
        triggered.add(61)
    if (z / (x / 50) > 0.9 and z / (x / 50) < 1.4) != (z / (x / 50) > 0.9 and z / (x / 50) < 133.4):
        triggered.add(62)

    if (x * y * z > 7000 and x * y * z < 11000) != (x * y * z > 7000 and x * y * 2 < 11000):
        triggered.add(63)
    if (x * y * z > 7000 and x * y * z < 11000) != (x * y * z > 7000 and x * y * 54 < 11000):
        triggered.add(64)

    if (x * 0.6 + y * 0.3 + z * 10 > 65 and x * 0.6 + y * 0.3 + z * 10 < 75) != (
        x * 0.6 + y * 0.3 + z * 10 > 645 and x * 0.6 + y * 0.3 + z * 10 < 75):
        triggered.add(65)
    if (x * 0.6 + y * 0.3 + z * 10 > 65 and x * 0.6 + y * 0.3 + z * 10 < 75) != (
        x * 0.6 + y * 40.3 + z * 10 > 65 and x * 0.6 + y * 0.3 + z * 10 < 75):
        triggered.add(66)

    if ((x / 65) ** 0.5 * (y / 100) ** 0.3 * (z / 1.35) ** 0.2 > 0.92) != (
        (x / 65) ** 0.5 * (y / 100) ** 0.3 * (z / 1.35) ** 0.2 > 392):
        triggered.add(67)
    if ((x / 65) ** 0.5 * (y / 100) ** 0.3 * (z / 1.35) ** 0.2 > 0.92) != (
        (x / 65) ** 0.5 * (y / 100) ** 0.3 * (z / 1.35) ** 0.2 > 982):
        triggered.add(68)

    if ((x - 65) * (y - 100) > -150 and (x - 65) * (y - 100) < 150) != (
        (x - 65) * (y - 100) > -150 and (x - 65) * (y - 100) < 1250):
        triggered.add(69)
    if ((x - 65) * (y - 100) > -150 and (x - 65) * (y - 100) < 150) != (
        (x - 65) * (y - 100) > -150 and (x - 65) * (y - 100) < 1050):
        triggered.add(70)

    if ((x - 65) * (z - 1.35) > -8 and (x - 65) * (z - 1.35) < 8) != (
        (x - 65) * (z - 1.35) > -8 and (x - 65) * (z - 1.35) < 18):
        triggered.add(71)
    if ((x - 65) * (z - 1.35) > -8 and (x - 65) * (z - 1.35) < 8) != (
        (x - 65) * (z - 1.35) > -8 and (x - 65) * (z - 1.35) < 448):
        triggered.add(72)

    if (abs(x / 65 + y / 100 + z / 1.35 - 3) < 0.25) != (abs(x / 65 + y / 100 + z / 1.35 - 3) < 25):
        triggered.add(73)
    if (abs(x / 65 + y / 100 + z / 1.35 - 3) < 0.25) != (abs(x / 65 + y / 100 + z / 1.35 - 3) < 40.25):
        triggered.add(74)

    if (abs((x / 65) * (y / 100) * (z / 1.35) - 1) < 0.18) != (abs((x / 65) * (y / 100) * (z / 1.35) - 1) < 40.18):
        triggered.add(75)
    if (abs((x / 65) * (y / 100) * (z / 1.35) - 1) < 0.18) != (abs((x / 65) * (y / 100) * (z / 1.35) - 1) < 18):
        triggered.add(76)

    if (x > 62 and x < 68 and y > 97 and y < 103 and z > 1.32 and z < 1.38) != (
        x > 62 and x < 68 and y > 97 and y < 103 and z > 1.32 and z < 138):
        triggered.add(77)
    if (x > 62 and x < 68 and y > 97 and y < 103 and z > 1.32 and z < 1.38) != (
        x > 62 and x < 68 and y > 97 and y < 103 and z > 1.32 and z < 228):
        triggered.add(78)

    if (abs(x / y - 0.65) < 0.08 and abs(z - 1.35) < 0.03) != (abs(x / y - 0.65) < 0.08 and abs(z - 1.35) < 3):
        triggered.add(79)
    if (abs(x / y - 0.65) < 0.08 and abs(z - 1.35) < 0.03) != (abs(x / y - 0.65) < 0.08 and abs(z - 1.35) < 40.03):
        triggered.add(80)

    if (x * y > 6300 and x * y < 6700 and z > 1.33 and z < 1.37) != (
        x * y > 6300 and x * y < 6700 and z > 1.33 and z < 41.37):
        triggered.add(81)
    if (x * y > 6300 and x * y < 6700 and z > 1.33 and z < 1.37) != (
        x * y > 6300 and x * y < 6700 and z > 1.33 and z < 91.37):
        triggered.add(82)

    if ((x + y) / 2 > 80 and (x + y) / 2 < 86 and z > 1.32 and z < 1.38) != (
        (x + y) / 2 > 80 and (x + y) / 2 < 86 and z > 1.32 and z < 138):
        triggered.add(83)
    if ((x + y) / 2 > 80 and (x + y) / 2 < 86 and z > 1.32 and z < 1.38) != (
        (x + y) / 2 > 80 and (x + y) / 2 < 86 and z > 1.32 and z < 19.38):
        triggered.add(84)

    if (abs(x - y) < 40 and z > 1.32 and z < 1.38) != (abs(x - y) < 40 and z > 1.32 and z < 138):
        triggered.add(85)
    if (abs(x - y) < 40 and z > 1.32 and z < 1.38) != (abs(x - y) < 40 and z > 1.32 and z < 13.8):
        triggered.add(86)

    if (math.sqrt((x - 65) ** 2 + (y - 100) ** 2) < 5 and abs(z - 1.35) < 0.03) != (
        math.sqrt((x - 65) ** 2 + (y - 100) ** 2) < 5 and abs(z - 1.35) < 13):
        triggered.add(87)
    if (math.sqrt((x - 65) ** 2 + (y - 100) ** 2) < 5 and abs(z - 1.35) < 0.03) != (
        math.sqrt((x - 65) ** 2 + (y - 100) ** 2) < 5 and abs(z - 1.35) < 8):
        triggered.add(88)

    if (x / 65 > 0.975 and x / 65 < 1.025 and y / 100 > 0.97 and y / 100 < 1.03) != (
        x / 65 > 0.975 and x / 65 < 1.025 and y / 100 > 10.97 and y / 100 < 1.03):
        triggered.add(89)
    if (x / 65 > 0.975 and x / 65 < 1.025 and y / 100 > 0.97 and y / 100 < 1.03) != (
        x / 65 > 0.975 and x / 65 < 1.025 and y / 100 > 40.97 and y / 100 < 1.03):
        triggered.add(90)

    if (z / 1.35 > 0.978 and z / 1.35 < 1.022) != (z / 1.35 > 0.978 and z / 1.35 < 41.022):
        triggered.add(91)
    if (z / 1.35 > 0.978 and z / 1.35 < 1.022) != (z / 1.35 > 0.978 and z / 1.35 < 622):
        triggered.add(92)

    if ((x / 65 + y / 100 + z / 1.35) / 3 > 0.975 and (x / 65 + y / 100 + z / 1.35) / 3 < 1.025) != (
        (x / 65 + y / 100 + z / 1.35) / 3 > 9.75 and (x / 65 + y / 100 + z / 1.35) / 3 < 1.025):
        triggered.add(93)
    if ((x / 65 + y / 100 + z / 1.35) / 3 > 0.975 and (x / 65 + y / 100 + z / 1.35) / 3 < 1.025) != (
        (x / 65 + y / 100 + z / 1.35) / 3 > 0.975 and (x / 65 + y / 100 + z / 1.35) / 3 < 10.25):
        triggered.add(94)

    if (max(abs(x / 65 - 1), abs(y / 100 - 1), abs(z / 1.35 - 1)) < 0.025) != (
        max(abs(x / 65 - 1), abs(y / 100 - 1), abs(z / 1.35 - 1)) < 40.025):
        triggered.add(95)
    if (max(abs(x / 65 - 1), abs(y / 100 - 1), abs(z / 1.35 - 1)) < 0.025) != (
        max(abs(x / 65 - 1), abs(y / 100 - 1), abs(z / 1.35 - 1)) < 10.025):
        triggered.add(96)

    if (min(x / 65, y / 100, z / 1.35) > 0.975) != (min(x / 65, y / 100, z / 1.35) > 9.75):
        triggered.add(97)
    if (min(x / 65, y / 100, z / 1.35) > 0.975) != (min(x / 65, y / 100, z / 1.35) > 19.75):
        triggered.add(98)

    if (max(x / 65, y / 100, z / 1.35) < 1.025) != (max(x / 65, y / 100, z / 1.35) < 11.025):
        triggered.add(99)
    if (max(x / 65, y / 100, z / 1.35) < 1.025) != (max(x / 65, y / 100, z / 1.35) < 41.025):
        triggered.add(100)

    if (abs(max(x, y, z * 100) - min(x, y, z * 100)) < 40) != (abs(max(x, y, z * 100) - min(x, y, z * 100)) < 140):
        triggered.add(101)
    if (abs(max(x, y, z * 100) - min(x, y, z * 100)) < 40) != (abs(max(x, y, z * 100) - min(x, y, z * 100)) < 1140):
        triggered.add(102)

    if ((x + y + z * 100) > 298 and (x + y + z * 100) < 302) != (
        (x + y + z * 100) > 298 and (x + y + z * 100) < 3102):
        triggered.add(103)
    if ((x + y + z * 100) > 298 and (x + y + z * 100) < 302) != (
        (x + y + z * 100) > 298 and (x + y + z * 100) < 352):
        triggered.add(104)

    if (x * y * z > 8700 and x * y * z < 9300) != (x * y * z > 8700 and x * y * 4 < 9300):
        triggered.add(105)
    if (x * y * z > 8700 and x * y * z < 9300) != (x * y * z > 8700 and x * y * z < 5300):
        triggered.add(106)

    if (abs((x + y + z * 100) / 3 - 100) < 2) != (abs((x + y + z * 100) / 3 - 100) < 332):
        triggered.add(107)
    if (abs((x + y + z * 100) / 3 - 100) < 2) != (abs((x + y + z * 100) / 3 - 100) < 122):
        triggered.add(108)

    if (z > 1.345 and z < 1.355) != (z > 1.345 and z < 1355):
        triggered.add(109)
    if (z > 1.345 and z < 1.355) != (z > 1.345 and z < 355):
        triggered.add(110)

    if (x < 52 or y < 85 or z < 1.22) != (x < 52 or y < 85 or z < 122):
        triggered.add(111)
    if (x < 52 or y < 85 or z < 1.22) != (x < 52 or y < 855 or z < 1.22):
        triggered.add(112)

    if (x > 78 or y > 115 or z > 1.48) != (x > 78 or y > 115 or z > 148):
        triggered.add(113)
    if (x > 78 or y > 115 or z > 1.48) != (x > 78 or y > 115 or z * 8 > 1.48):
        triggered.add(114)

    if (abs(x / y - 0.65) > 0.12) != (abs(x / y - 0.65) > 12):
        triggered.add(115)
    if (abs(x / y - 0.65) > 0.12) != (abs(x / y - 0.65) > 20.12):
        triggered.add(116)

    if (abs(x / z - 48) > 6) != (abs(x / z - 48) > 16):
        triggered.add(117)
    if (abs(x / z - 48) > 6) != (abs(x / z - 48) > 36):
        triggered.add(118)

    if (abs(y / z - 74) > 8) != (abs(y / z - 74) > 228):
        triggered.add(119)
    if (abs(y / z - 74) > 8) != (abs(y / z - 74) > 82):
        triggered.add(120)

    if ((x + y + z * 100) < 285 or (x + y + z * 100) > 315) != (
        (x + y + z * 100) < 285 or (x + y + z * 100) > 3165):
        triggered.add(121)
    if ((x + y + z * 100) < 285 or (x + y + z * 100) > 315) != (
        (x + y + z * 100) < 2835 or (x + y + z * 100) > 315):
        triggered.add(122)

    if (x * y * z < 8000 or x * y * z > 10000) != (x * y * z < 8000 or x * y * z > 1000):
        triggered.add(123)
    if (x * y * z < 8000 or x * y * z > 10000) != (x * y * z < 8000 or x * y * z > 93000):
        triggered.add(124)

    if (x < 55 and y < 90) != (x < 55 and y < 290):
        triggered.add(125)
    if (x < 55 and y < 90) != (x < 55 and 80 < 90):
        triggered.add(126)

    if (x > 75 and y > 110) != (x > 75 and 200 > 110):
        triggered.add(127)
    if (x > 75 and y > 110) != (x > 75 and 500 > 110):
        triggered.add(128)

    if (x < 55 and z < 1.25) != (x < 55 and z < 125):
        triggered.add(129)
    if (x < 55 and z < 1.25) != (x < 55 and z < 12.5):
        triggered.add(130)

    if (x > 75 and z > 1.45) != (x > 75 and z > 145):
        triggered.add(131)
    if (x > 75 and z > 1.45) != (x > 75 and 6 > 1.45):
        triggered.add(132)

    if (y < 90 and z < 1.25) != (y < 90 and 1 < 1.25):
        triggered.add(133)
    if (y < 90 and z < 1.25) != (y < 90 and z < 125):
        triggered.add(134)

    if (y < 75 or y > 125) != (y < 75 or y * 8 > 125):
        triggered.add(135)
    if (y < 75 or y > 125) != (y < 75 or y * 10 > 125):
        triggered.add(136)

    if (x < 45 and y < 80 and z < 1.2) != (x < 45 and y < 80 and z < 12):
        triggered.add(137)
    if (x < 45 and y < 80 and z < 1.2) != (x < 45 and y < 80 and z < 11.2):
        triggered.add(138)

    if (abs((x * y * z) / 9000 - 1) > 0.22) != (abs((x * y * z) / 9000 - 1) > 9.22):
        triggered.add(139)
    if (abs((x * y * z) / 9000 - 1) > 0.22) != (abs((x * y * z) / 9000 - 1) > 22):
        triggered.add(140)

    return triggered

# === 目标路径组（section7 的编号 1~140） ===
targetPaths = [
    {1, 3, 5, 6, 7, 9, 10, 11, 12, 13, 17, 18, 19, 20, 23, 24, 25, 27, 28, 30, 32, 33, 34, 36, 39, 47, 48, 49, 50, 51,
     52, 55, 56, 57, 58, 61, 62, 67, 68, 73, 74, 75, 76, 79, 80, 83, 84, 85, 86, 87, 88, 91, 92, 94, 95, 96, 97, 98, 99,
     100, 101, 102, 103, 107, 108, 109, 110, 111, 112, 113, 117, 118, 119, 120, 121, 124, 135, 136, 139, 140},

    {1, 3, 5, 6, 7, 9, 10, 11, 12, 13, 15, 16, 17, 18, 19, 20, 23, 24, 25, 27, 28, 30, 32, 33, 34, 36, 39, 49, 50, 51,
     52, 55, 56, 57, 58, 61, 62, 67, 68, 73, 74, 75, 76, 79, 80, 83, 84, 85, 86, 87, 88, 91, 92, 94, 95, 96, 97, 98, 99,
     100, 101, 102, 103, 107, 108, 109, 110, 111, 112, 113, 117, 118, 119, 120, 121, 124, 135, 136, 139, 140},

    {1, 3, 5, 6, 9, 10, 11, 12, 13, 17, 18, 19, 20, 21, 22, 23, 24, 25, 27, 28, 30, 32, 33, 34, 36, 39, 49, 50, 51, 52,
     53, 54, 55, 56, 57, 58, 61, 62, 67, 68, 73, 74, 75, 76, 81, 82, 83, 84, 85, 86, 91, 92, 94, 95, 96, 99, 100, 101,
     102, 103, 107, 108, 109, 110, 111, 112, 113, 115, 116, 117, 118, 119, 120, 121, 124, 135, 136, 139, 140},

    {2, 3, 5, 6, 11, 12, 13, 17, 18, 20, 21, 22, 23, 24, 25, 30, 32, 33, 34, 36, 39, 41, 42, 49, 50, 51, 52, 54, 55, 56,
     57, 58, 67, 68, 71, 72, 73, 74, 75, 76, 83, 84, 85, 86, 91, 92, 94, 95, 96, 99, 100, 101, 102, 103, 107, 108, 109,
     110, 111, 112, 115, 116, 117, 118, 119, 120, 121, 124, 127, 128, 131, 133, 134, 135, 136, 139, 140},

    {1, 3, 5, 6, 8, 9, 10, 11, 12, 13, 17, 18, 19, 20, 21, 22, 23, 24, 25, 27, 28, 30, 32, 33, 34, 36, 39, 49, 50, 51,
     52, 53, 54, 55, 56, 57, 58, 67, 68, 73, 74, 75, 76, 81, 82, 83, 84, 85, 86, 91, 92, 94, 95, 96, 99, 100, 101, 102,
     103, 107, 108, 109, 110, 111, 112, 113, 115, 116, 117, 118, 119, 120, 121, 124, 135, 136, 139, 140},

    {1, 3, 5, 6, 7, 9, 10, 11, 12, 15, 16, 17, 18, 23, 24, 25, 27, 28, 30, 32, 33, 34, 36, 39, 49, 50, 51, 52, 55, 56,
     57, 58, 61, 62, 67, 68, 73, 74, 75, 76, 77, 78, 79, 80, 83, 84, 85, 86, 87, 88, 91, 92, 94, 95, 96, 99, 100, 101,
     102, 103, 107, 108, 109, 110, 111, 112, 113, 118, 119, 120, 121, 124, 135, 136, 139, 140},

    {1, 3, 5, 6, 9, 10, 11, 12, 14, 17, 18, 21, 22, 23, 24, 25, 27, 28, 30, 32, 33, 34, 36, 39, 47, 48, 49, 50, 51, 52,
     53, 54, 55, 56, 57, 58, 61, 62, 67, 68, 73, 74, 75, 76, 79, 80, 91, 92, 94, 95, 96, 97, 98, 99, 100, 101, 102, 103,
     107, 108, 109, 110, 111, 112, 113, 117, 118, 119, 120, 121, 124, 135, 136, 139, 140},

    {1, 3, 5, 6, 9, 10, 13, 17, 18, 19, 20, 22, 23, 24, 27, 28, 30, 32, 33, 34, 36, 39, 41, 42, 49, 50, 51, 52, 53, 54,
     55, 56, 57, 58, 61, 62, 67, 68, 73, 74, 75, 76, 85, 86, 91, 92, 94, 95, 96, 99, 100, 101, 102, 103, 104, 107, 108,
     109, 110, 113, 115, 116, 117, 118, 119, 120, 121, 124, 133, 134, 135, 136, 139, 140},

    {1, 4, 5, 6, 13, 17, 18, 20, 22, 23, 24, 30, 32, 33, 34, 36, 39, 41, 42, 49, 50, 51, 52, 54, 55, 56, 57, 58, 59, 60,
     67, 68, 71, 72, 73, 74, 75, 76, 85, 86, 91, 92, 94, 95, 96, 99, 100, 101, 102, 103, 107, 108, 109, 110, 113, 115,
     116, 117, 118, 119, 120, 121, 124, 127, 128, 131, 133, 134, 135, 136, 139, 140},

    {1, 3, 5, 6, 11, 12, 13, 17, 18, 20, 21, 22, 23, 24, 25, 30, 33, 34, 36, 39, 47, 48, 49, 50, 51, 52, 55, 56, 57, 58,
     67, 68, 69, 70, 71, 72, 73, 74, 75, 76, 79, 80, 85, 86, 91, 92, 94, 95, 96, 97, 98, 99, 100, 101, 102, 103, 107,
     108, 109, 110, 111, 112, 117, 118, 119, 120, 121, 124, 131, 135, 136, 139, 140},

    {1, 4, 5, 6, 9, 10, 13, 17, 18, 19, 20, 22, 23, 24, 27, 28, 30, 31, 33, 34, 36, 39, 41, 42, 49, 50, 51, 52, 54, 55,
     56, 57, 58, 61, 62, 64, 67, 68, 73, 74, 75, 76, 85, 86, 91, 92, 94, 95, 96, 99, 100, 101, 102, 103, 104, 107, 108,
     109, 110, 113, 115, 116, 117, 118, 119, 120, 121, 124, 133, 134, 135, 136},
    {1, 3, 5, 7, 9, 10, 11, 12, 14, 15, 16, 17, 18, 23, 25, 27, 28, 32, 33, 34, 36, 39, 50, 52, 61, 62, 67, 68, 73, 74,
     75, 76, 77, 78, 80, 81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 94, 95, 96, 97, 98, 99, 100, 102, 103, 107,
     108, 109, 110, 111, 112, 113, 118, 119, 120, 121, 124, 135, 136, 139, 140},

    {1, 4, 5, 6, 13, 17, 18, 19, 20, 22, 23, 24, 30, 32, 33, 34, 35, 39, 41, 42, 49, 50, 51, 52, 54, 55, 56, 57, 58, 59,
     60, 67, 68, 73, 74, 75, 76, 85, 86, 91, 92, 94, 95, 96, 99, 100, 101, 102, 103, 104, 107, 108, 109, 110, 113, 115,
     116, 117, 118, 119, 120, 121, 124, 127, 128, 131, 133, 134, 139, 140},
    {1, 3, 5, 6, 11, 12, 17, 18, 20, 21, 22, 23, 24, 27, 28, 30, 32, 36, 39, 49, 50, 51, 52, 54, 55, 56, 57, 58, 61, 62,
     67, 68, 73, 74, 75, 76, 83, 84, 91, 92, 94, 95, 96, 99, 100, 102, 103, 107, 108, 109, 110, 111, 112, 113, 115, 116,
     118, 119, 120, 121, 124, 125, 126, 129, 130, 135, 136, 139, 140},

    {1, 3, 5, 6, 11, 12, 17, 18, 19, 20, 21, 22, 23, 24, 27, 28, 30, 31, 36, 39, 49, 50, 51, 52, 55, 56, 57, 58, 61, 62,
     64, 67, 68, 69, 70, 79, 80, 85, 86, 91, 92, 94, 95, 96, 99, 100, 102, 103, 104, 105, 106, 107, 108, 109, 110, 111,
     112, 113, 118, 119, 120, 121, 123, 129, 130, 133, 134, 135, 136},

    {1, 4, 5, 6, 9, 10, 13, 17, 18, 19, 20, 22, 23, 24, 27, 28, 30, 31, 33, 34, 36, 39, 41, 42, 50, 51, 52, 54, 55, 56,
     59, 60, 61, 62, 63, 67, 68, 71, 72, 73, 74, 75, 76, 85, 86, 91, 92, 94, 95, 96, 99, 100, 102, 103, 107, 108, 109,
     110, 113, 115, 116, 118, 119, 120, 121, 124, 133, 134, 139, 140},

    {1, 4, 5, 6, 17, 18, 19, 20, 22, 23, 24, 27, 28, 29, 31, 35, 37, 38, 39, 41, 42, 49, 50, 51, 52, 53, 54, 55, 56, 57,
     58, 61, 62, 65, 69, 70, 75, 76, 85, 86, 91, 92, 93, 95, 96, 99, 100, 102, 103, 104, 107, 108, 109, 110, 113, 115,
     116, 118, 119, 120, 121, 129, 130, 133, 134, 139, 140},

    {2, 4, 5, 6, 17, 18, 22, 23, 24, 27, 28, 29, 31, 36, 37, 38, 39, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 61, 62, 66,
     69, 70, 75, 76, 85, 86, 91, 92, 95, 96, 99, 100, 102, 103, 104, 107, 108, 109, 110, 113, 115, 116, 118, 119, 120,
     122, 129, 130, 133, 134, 135, 136, 137, 138, 139, 140},

    {1, 4, 13, 19, 20, 22, 23, 24, 26, 29, 35, 37, 38, 39, 43, 45, 47, 48, 55, 56, 57, 58, 63, 64, 67, 68, 69, 70, 79,
     80, 94, 95, 96, 99, 100, 101, 102, 105, 106, 107, 108, 118, 119, 120, 122, 123, 135, 136},

    {1, 3, 13, 20, 21, 22, 23, 24, 26, 29, 35, 37, 38, 39, 43, 45, 47, 48, 55, 56, 57, 58, 63, 64, 67, 68, 69, 70, 79,
     80, 93, 95, 96, 99, 100, 105, 106, 107, 108, 118, 119, 120, 122, 123, 132, 135, 136},

    {1, 3, 8, 13, 19, 20, 21, 22, 23, 24, 26, 29, 37, 38, 39, 43, 45, 47, 48, 55, 56, 57, 58, 63, 64, 67, 68, 79, 80,
     93, 95, 96, 99, 100, 107, 108, 114, 118, 119, 120, 123, 127, 128, 132, 135, 136},

    {2, 3, 21, 22, 23, 24, 31, 40, 45, 54, 55, 56, 57, 58, 61, 62, 66, 71, 72, 73, 74, 75, 76, 95, 96, 99, 100, 101,
     102, 107, 108, 114, 115, 116, 117, 118, 119, 120, 125, 126, 135, 136, 139, 140},

    {2, 3, 20, 21, 22, 23, 24, 31, 40, 44, 45, 53, 54, 57, 58, 66, 69, 70, 71, 72, 73, 74, 75, 76, 95, 96, 101, 102,
     107, 108, 114, 115, 116, 117, 118, 119, 120, 125, 126, 135, 136, 139, 140},

    {2, 3, 21, 22, 23, 24, 31, 40, 46, 54, 55, 56, 57, 58, 61, 62, 66, 69, 70, 71, 72, 73, 74, 75, 76, 95, 96, 101, 102,
     107, 108, 114, 115, 116, 117, 118, 119, 120, 135, 136, 139, 140},
]

# 将路径列表转换为集合列表
target_paths = [set(path) for path in targetPaths]
NUM_PATHS = len(target_paths)

def jaccard_similarity(set1, set2):
    """Compute Jaccard similarity"""
    if not set2:
        return 0.0

    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0

# === Path Similarity ===
def compute_path_similarity_matrix(paths):
    """Path Similarity"""
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix

def group_paths_by_similarity(paths, threshold_percentile=50):
    """SimilarityPath """
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.percentile(avg_sim_scores, threshold_percentile)

    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]

    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)

    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group

# === Sample generation ===
def compute_robustness(state, path, sample_size=9):
    """计算鲁棒性"""
    base = section7_flow_pressure_density_hybrid(state[0], state[1], state[2])
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0

    # 邻域偏移量（z范围较小，步长适当缩小）
    deltas = [
        (-1, -1, -3), (0, -1, 0), (1, -1, 3),
        (-1, 0, -3), (1, 0, 3),
        (-1, 1, -3), (0, 1, 0), (1, 1, 3),
        (0, 0, 0)
    ]

    for dw, dt, dz in deltas[:sample_size]:
        if dw == dt == dz == 0:
            continue

        neighbor_weather = int(np.clip(state[0] + dw, MIN_X, MAX_X))
        neighbor_time = int(np.clip(state[1] + dt, MIN_Y, MAX_Y))
        neighbor_z = int(np.clip(state[2] + dz, MIN_Z, MAX_Z))
        neighbor = (neighbor_weather, neighbor_time, neighbor_z)

        n_trig = section7_flow_pressure_density_hybrid(neighbor[0], neighbor[1], neighbor[2])
        if not n_trig:
            continue

        rob += jaccard_similarity(n_trig, base)
        neighbors += 1

    return rob / neighbors if neighbors > 0 else 0.0

def generate_samples_for_all_paths(num_candidates=2000, top_k=200, run_id=1):
    """为所有路径生成样本（权重筛选）"""
    BEST_WEIGHTS = [0.55, 0.25, 0.2]

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_individual.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Individual Path {path_id} (Weighted Screening) - Run {run_id}\n")
            f.write("weather time_period z\tScore\tSimilarity\tRobustness\tLengthDiff\n")
            for s in samples:
                weather, time_period, z = s['state']
                f.write(
                    f"{weather} {time_period} {z}\t{s['score']:.4f}\t{s['similarity']:.4f}\t"
                    f"{s['robustness']:.4f}\t{s['length_diff']:.4f}\n"
                )

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_individual"

    for path_idx in range(len(target_paths)):
        path = target_paths[path_idx]
        candidate_samples = []
        attempts = 0
        max_attempts = num_candidates * 10

        while len(candidate_samples) < num_candidates and attempts < max_attempts:
            attempts += 1

            weather = np.random.randint(MIN_X, MAX_X + 1)
            time_period = np.random.randint(MIN_Y, MAX_Y + 1)
            z = np.random.randint(MIN_Z, MAX_Z + 1)
            state = (weather, time_period, z)

            triggered = section7_flow_pressure_density_hybrid(weather, time_period, z)
            if not triggered:
                continue

            sim = jaccard_similarity(triggered, path)
            rob = compute_robustness(state, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))

            candidate_samples.append({
                'state': state,
                'similarity': sim,
                'robustness': rob,
                'length_diff': len_diff,
                'triggered': triggered
            })

        if candidate_samples:
            for sample in candidate_samples:
                score = (BEST_WEIGHTS[0] * sample['similarity'] +
                         BEST_WEIGHTS[1] * sample['robustness'] +
                         BEST_WEIGHTS[2] * sample['length_diff'])
                sample['score'] = score

            candidate_samples.sort(key=lambda x: x['score'], reverse=True)
            selected_samples = candidate_samples[:top_k]
            save_samples(path_id=path_idx + 1, samples=selected_samples, base_dir=base_dir)

# === Shared Experience Replay ===
class SharedExperienceReplay:
    def __init__(self, capacity=20000):
        self.capacity = capacity
        self.buffer = deque(maxlen=self.capacity)
        self.priorities = deque(maxlen=self.capacity)

    def append(self, experience):
        self.buffer.append(experience)
        self.priorities.append(experience[-1])

    def sample(self, batch_size, alpha=0.6):
        if len(self.buffer) < batch_size:
            return [], [], []

        priorities = np.array(self.priorities, dtype=np.float64)
        priorities = np.power(priorities, alpha)
        probabilities = priorities / np.sum(priorities)

        batch_indices = np.random.choice(len(self.buffer), batch_size, p=probabilities, replace=False)
        batch = [self.buffer[idx] for idx in batch_indices]

        return batch, batch_indices, probabilities[batch_indices]

    def update_priorities(self, batch_indices, td_errors):
        for idx, td_error in zip(batch_indices, td_errors):
            if idx < len(self.priorities):
                self.priorities[idx] = max(abs(td_error), 1e-6)

    def __len__(self):
        return len(self.buffer)

    def get_high_reward_samples(self, target_path, num_samples=20):
        if len(self.buffer) == 0:
            return []

        samples_with_scores = []
        seen_states = set()

        for experience in self.buffer:
            state_tensor = experience[0]
            state_norm = state_tensor.cpu().numpy().flatten()
            state_tuple = denormalize_state((state_norm[0], state_norm[1], state_norm[2]))

            if state_tuple in seen_states:
                continue
            seen_states.add(state_tuple)

            triggered = section7_flow_pressure_density_hybrid(state_tuple[0], state_tuple[1], state_tuple[2])
            reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)

            samples_with_scores.append((state_tuple, reward, sim, triggered))

        samples_with_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_scores[:num_samples]

def load_path_data(file_path):
    path_data = []

    if not os.path.exists(file_path):
        print(f"文件不存在: {file_path}")
        return path_data

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
            for line in lines[2:]:
                parts = line.strip().split("\t")
                if parts:
                    values = parts[0].split()
                    if len(values) >= 3:
                        state = (int(values[0]), int(values[1]), int(values[2]))
                        path_data.append(state)
    except Exception as e:
        print(f"读取文件失败 {file_path}: {e}")

    return path_data

# === DQN ===
class DQN(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dims=[128, 64]):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, hidden_dims[0])
        self.fc2 = nn.Linear(hidden_dims[0], hidden_dims[1])
        self.fc3 = nn.Linear(hidden_dims[1], action_dim)
        self.dropout = nn.Dropout(0.1)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = self.dropout(x)
        x = torch.relu(self.fc2(x))
        return self.fc3(x)

# === DQN Agent with PER ===
class DQNAgentWithPER:
    def __init__(self, state_dim, action_dim, replay_buffer,
                 gamma=0.99, epsilon=1.0, epsilon_decay=0.995,
                 epsilon_min=0.1, learning_rate=0.001, alpha=0.6, beta=0.4):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer
        self.alpha = alpha
        self.beta = beta

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        """
        动作解码：30个动作，分别对应三个维度的不同步长
        - weather: +/-1, 0(x2)
        - time_period: +/-1, 0(x2)
        - z: +/-3, +/-2, +/-1, 0(x2)  适配 1~15 范围
        """
        delta_values_weather_time = [1, 0, 0, -1]
        delta_values_z = [3, 2, 1, 0, 0, -1, -2, -3]

        dim = action_idx // 10
        delta_idx = action_idx % 10

        if dim == 0:  # weather
            if delta_idx >= 4:
                delta_idx = 3
            return (delta_values_weather_time[delta_idx], 0, 0)
        elif dim == 1:  # time_period
            if delta_idx >= 4:
                delta_idx = 3
            return (0, delta_values_weather_time[delta_idx], 0)
        elif dim == 2:  # z
            if delta_idx >= 8:
                delta_idx = 7
            return (0, 0, delta_values_z[delta_idx])

    def act(self, state_norm, legal_actions=None):
        if legal_actions is None:
            legal_actions = list(range(self.action_dim))

        if not legal_actions:
            return None

        if random.random() < self.epsilon:
            return random.choice(legal_actions)

        state_tensor = torch.tensor(state_norm, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)[0]

        legal_q_values = q_values[legal_actions]
        best_legal_idx = torch.argmax(legal_q_values).item()
        return legal_actions[best_legal_idx]

    def get_legal_actions(self, state):
        legal_actions = []

        for action_idx in range(self.action_dim):
            dw, dt, dz = self.decode_action(action_idx)

            next_weather = state[0] + dw
            next_time = state[1] + dt
            next_z = state[2] + dz

            if (MIN_X <= next_weather <= MAX_X and
                    MIN_Y <= next_time <= MAX_Y and
                    MIN_Z <= next_z <= MAX_Z):
                legal_actions.append(action_idx)

        return legal_actions

    def store_transition(self, state, action, reward, next_state, done):
        state_norm = normalize_state(state)
        next_state_norm = normalize_state(next_state)

        state_tensor = torch.tensor(state_norm, dtype=torch.float32).unsqueeze(0).to(device)
        next_state_tensor = torch.tensor(next_state_norm, dtype=torch.float32).unsqueeze(0).to(device)

        with torch.no_grad():
            q_values = self.model(state_tensor)
            next_q_values = self.target_model(next_state_tensor)
            max_next_q = next_q_values.max(1)[0]
            target_q = reward + (self.gamma * max_next_q * (1 - done))
            td_error = abs(q_values[0][action].item() - target_q.item())

        self.replay_buffer.append((state_tensor, action, reward, next_state_tensor, done, td_error))
        return td_error

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return 0.0

        batch, batch_indices, probabilities = self.replay_buffer.sample(batch_size, alpha=self.alpha)

        if not batch:
            return 0.0

        states, actions, rewards, next_states, dones, _ = zip(*batch)

        weights = (len(self.replay_buffer) * probabilities) ** (-self.beta)
        weights = weights / weights.max()
        weights = torch.tensor(weights, dtype=torch.float32).to(device)

        states = torch.cat(states).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.cat(next_states).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q = self.target_model(next_states).max(1)[0].detach()
        target_q = rewards + (self.gamma * next_max_q * (1 - dones))

        td_errors = current_q - target_q
        weighted_loss = (td_errors.pow(2) * weights).mean()

        self.optimizer.zero_grad()
        weighted_loss.backward()
        torch.nn.utils.clip_grad_norm_(self.model.parameters(), 1.0)
        self.optimizer.step()

        new_priorities = abs(td_errors.detach().cpu().numpy())
        self.replay_buffer.update_priorities(batch_indices, new_priorities)

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

        return weighted_loss.item()

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())

# === 训练函数 ===
def generate_and_train_for_individual_paths(path_documents, repeats=5, batch_size=32, run_id=1):
    state_dim = 3
    action_dim = 30

    shared_replay_buffer = SharedExperienceReplay(capacity=20000)
    agent = DQNAgentWithPER(state_dim, action_dim, shared_replay_buffer)

    total_cumulative_reward = 0
    path_rewards = {}

    print(f"\n=== 第 {run_id}/20 次训练开始 ===")
    start_time = time.time()

    SAMPLES_PER_BATCH = 50
    NUM_BATCHES = 4
    STEPS_PER_SAMPLE = 3

    for path_idx in range(len(target_paths)):
        path_id = path_idx + 1
        print(f"\n{'=' * 60}")
        print(f"路径 {path_id}/{NUM_PATHS}")
        print(f"{'=' * 60}")

        file_path = os.path.join(path_documents, f"path{path_id}_individual.txt")
        if not os.path.exists(file_path):
            print(f"  文件不存在: {file_path}")
            continue

        path_data = load_path_data(file_path)
        if not path_data:
            print(f"  路径 {path_id} 无有效数据")
            continue

        target_path = target_paths[path_idx]

        if path_idx not in path_rewards:
            path_rewards[path_idx] = 0

        for repeat_idx in range(repeats):
            print(f"\n  重复 {repeat_idx + 1}/{repeats}")

            for batch_idx in range(NUM_BATCHES):
                batch_start = batch_idx * SAMPLES_PER_BATCH
                batch_end = min(batch_start + SAMPLES_PER_BATCH, len(path_data))

                print(f"     批次 {batch_idx + 1}/{NUM_BATCHES} (样本 {batch_start}-{batch_end})")

                for sample_idx in range(batch_start, batch_end):
                    state = path_data[sample_idx]
                    prev_state = None
                    prev_triggered = None

                    for step in range(STEPS_PER_SAMPLE):
                        legal_actions = agent.get_legal_actions(state)

                        if not legal_actions:
                            break

                        state_norm = normalize_state(state)
                        action = agent.act(state_norm, legal_actions)
                        if action is None:
                            break

                        dw, dt, dz = agent.decode_action(action)

                        next_state = (
                            int(np.clip(state[0] + dw, MIN_X, MAX_X)),
                            int(np.clip(state[1] + dt, MIN_Y, MAX_Y)),
                            int(np.clip(state[2] + dz, MIN_Z, MAX_Z))
                        )

                        triggered = section7_flow_pressure_density_hybrid(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered,
                                                prev_triggered, prev_state)
                        done = (step == STEPS_PER_SAMPLE - 1)

                        agent.store_transition(state, action, reward, next_state, done)

                        prev_state = state
                        prev_triggered = triggered
                        state = next_state

                        total_cumulative_reward += reward
                        path_rewards[path_idx] += reward

                if len(agent.replay_buffer) >= batch_size:
                    loss = agent.train(batch_size)
                    print(f"         批次 {batch_idx + 1} 完成，损失: {loss:.4f}")

                if (batch_idx + 1) % 2 == 0:
                    agent.update_target_model()
                    print(f"         目标网络已更新 (批次 {batch_idx + 1})")

        print(f"\n路径 {path_id} 完成，累计奖励: {path_rewards[path_idx]:.2f}")
        print(f"共享经验池大小: {len(shared_replay_buffer)}")

    training_time = time.time() - start_time
    print(f"\n=== 第 {run_id}/20 次训练完成，用时: {training_time:.2f} 秒 ===")

    return agent, shared_replay_buffer, total_cumulative_reward, path_rewards, training_time

# === Excel 报告生成 ===
def generate_excel_report(all_runs, similar_group, isolated_group, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    sim_paths = [i+1 for i in similar_group]
    iso_paths = [i+1 for i in isolated_group]

    wb = Workbook()
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    h_color = "4472C4"
    sim_color = "E2EFDA"
    iso_color = "FCE4D6"
    s_color = "FFF2CC"

    # Sheet1: 路径相似度
    ws1 = wb.active
    ws1.title = "路径相似度"
    headers1 = ['Path ID', '分组'] + [f'Run {i}' for i in range(1, 21)] + ['平均', '最高', '最低', '标准差']
    for c, h in enumerate(headers1, 1):
        cell = ws1.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=h_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for pid in range(1, NUM_PATHS + 1):
        row = pid + 1
        if pid in sim_paths:
            gtype, gcolor = "高相似组", sim_color
        elif pid in iso_paths:
            gtype, gcolor = "低相似组", iso_color
        else:
            gtype, gcolor = "未分组", "FFFFFF"

        ws1.cell(row, 1, f"Path {pid}").font = Font(bold=True)
        ws1.cell(row, 2, gtype)
        for c in [1, 2]:
            ws1.cell(row, c).fill = PatternFill("solid", fgColor=gcolor)
            ws1.cell(row, c).alignment = Alignment(horizontal="center")
            ws1.cell(row, c).border = border

        sims = []
        for ri, run in enumerate(all_runs):
            s = run['path_sims'].get(pid, {}).get('avg', 0.0)
            sims.append(s)
            cell = ws1.cell(row, 3 + ri, round(s, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        stats = [np.mean(sims), np.max(sims), np.min(sims), np.std(sims)]
        for i, v in enumerate(stats):
            cell = ws1.cell(row, 23 + i, round(v, 4))
            cell.number_format = '0.0000'
            cell.fill = PatternFill("solid", fgColor=s_color)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    # Sheet2: 分组统计
    ws2 = wb.create_sheet("分组统计")
    headers2 = ['分组', '包含路径'] + [f'Run {i}' for i in range(1, 21)] + ['平均相似度', '标准差']
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=h_color)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    def write_group_row(row, name, paths, color):
        ws2.cell(row, 1, name).font = Font(bold=True)
        ws2.cell(row, 2, ','.join(map(str, paths)))
        for c in [1, 2]:
            ws2.cell(row, c).fill = PatternFill("solid", fgColor=color)
            ws2.cell(row, c).alignment = Alignment(horizontal="center")
            ws2.cell(row, c).border = border
        vals = []
        for ri, run in enumerate(all_runs):
            v = np.mean([run['path_sims'].get(p, {}).get('avg', 0.0) for p in paths])
            vals.append(v)
            cell = ws2.cell(row, 3 + ri, round(v, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        cell = ws2.cell(row, 23, round(np.mean(vals), 4))
        cell.number_format = '0.0000'
        cell.fill = PatternFill("solid", fgColor=s_color)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        cell = ws2.cell(row, 24, round(np.std(vals), 4))
        cell.number_format = '0.0000'
        cell.fill = PatternFill("solid", fgColor=s_color)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    write_group_row(2, "高相似组", sim_paths, sim_color)
    write_group_row(3, "低相似组", iso_paths, iso_color)

    # Sheet3: 轮次汇总
    ws3 = wb.create_sheet("轮次汇总")
    headers3 = ['轮次', '耗时(秒)', '总体平均相似度', '最高相似度', '最低相似度', '高相似组平均', '低相似组平均', '回放池容量']
    for c, h in enumerate(headers3, 1):
        cell = ws3.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=h_color)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for ri, run in enumerate(all_runs, 1):
        row = ri + 1
        high_avg = np.mean([run['path_sims'].get(p, {}).get('avg', 0.0) for p in sim_paths])
        low_avg = np.mean([run['path_sims'].get(p, {}).get('avg', 0.0) for p in iso_paths]) if iso_paths else 0.0
        vals = [
            f"Run {ri}",
            round(run['time'], 2),
            round(run['overall_avg'], 4),
            round(run['max_sim'], 4),
            round(run['min_sim'], 4),
            round(high_avg, 4),
            round(low_avg, 4),
            20000
        ]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row, c, v)
            if c == 1:
                cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    # Sheet4: Top样本
    ws4 = wb.create_sheet("Top样本详情")
    headers4 = ['轮次', '路径', '序号', 'X', 'Y', 'Z', '相似度', '触发规则']
    for c, h in enumerate(headers4, 1):
        cell = ws4.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=h_color)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    r_idx = 2
    for ri, run in enumerate(all_runs, 1):
        for pid in range(1, NUM_PATHS + 1):
            samples = run['samples'].get(pid, [])
            pcolor = sim_color if pid in sim_paths else (iso_color if pid in iso_paths else "FFFFFF")
            for si, (st, _, sim, trig) in enumerate(samples, 1):
                x, y, z = st
                ws4.cell(r_idx, 1, f"Run {ri}").fill = PatternFill("solid", fgColor=pcolor)
                ws4.cell(r_idx, 2, f"Path {pid}").fill = PatternFill("solid", fgColor=pcolor)
                ws4.cell(r_idx, 3, si)
                ws4.cell(r_idx, 4, x)
                ws4.cell(r_idx, 5, y)
                ws4.cell(r_idx, 6, z)
                ws4.cell(r_idx, 7, round(sim, 4)).number_format = '0.0000'
                ws4.cell(r_idx, 8, ','.join(map(str, sorted(trig))))
                for c in range(1, 9):
                    ws4.cell(r_idx, c).alignment = Alignment(horizontal="center")
                    ws4.cell(r_idx, c).border = border
                r_idx += 1

    out_path = os.path.join(out_dir, "20轮训练汇总报告.xlsx")
    wb.save(out_path)
    print(f"Excel 报告已保存: {out_path}")

# === 20 轮总入口 ===
def run_20_times_training():
    model_dir = r"D:\Experiment\CNN\DQNNEW\saved_models_new_vars"
    sample_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_individual"
    report_dir = r"D:\Experiment\CNN\ComparisonExperiment2\excel_reports_new_vars"

    os.makedirs(model_dir, exist_ok=True)
    os.makedirs(sample_dir, exist_ok=True)
    os.makedirs(report_dir, exist_ok=True)

    similar_group, isolated_group = group_paths_by_similarity(target_paths)

    print("=" * 60)
    print("DQN 路径覆盖优化 - 20轮完整训练")
    print(f"参数范围: X [{MIN_X},{MAX_X}] | Y [{MIN_Y},{MAX_Y}] | Z [{MIN_Z},{MAX_Z}]")
    print(f"路径总数: {NUM_PATHS}")
    print(f"高相似组: {[i+1 for i in similar_group]}")
    print(f"低相似组: {[i+1 for i in isolated_group]}")
    print("=" * 60)

    all_runs = []

    for run_id in range(1, 21):
        print(f"\n===== 第 {run_id}/20 轮 =====")
        print("[1/2] 生成初始样本...")
        generate_samples_for_all_paths(num_candidates=2000, top_k=200, run_id=run_id)

        print("[2/2] 开始训练...")
        agent, buffer, total_r, path_r, elapsed = generate_and_train_for_individual_paths(
            sample_dir, repeats=5, batch_size=32, run_id=run_id
        )

        # 保存模型
        model_path = os.path.join(model_dir, f"model_run_{run_id}.pth")
        torch.save({
            'model_state_dict': agent.model.state_dict(),
            'optimizer_state_dict': agent.optimizer.state_dict(),
            'epsilon': agent.epsilon,
            'ranges': {'x': [MIN_X, MAX_X], 'y': [MIN_Y, MAX_Y], 'z': [MIN_Z, MAX_Z]}
        }, model_path)

        # 收集统计
        run_data = {'path_sims': {}, 'samples': {}, 'time': elapsed}
        all_sims = []
        for pi in range(NUM_PATHS):
            pid = pi + 1
            samples = buffer.get_high_reward_samples(target_paths[pi], 20)
            if samples:
                sims = [s[2] for s in samples]
                run_data['path_sims'][pid] = {
                    'avg': np.mean(sims),
                    'max': np.max(sims),
                    'min': np.min(sims)
                }
                run_data['samples'][pid] = samples
                all_sims.extend(sims)
            else:
                run_data['path_sims'][pid] = {'avg': 0.0, 'max': 0.0, 'min': 0.0}
                run_data['samples'][pid] = []

        run_data['overall_avg'] = np.mean(all_sims) if all_sims else 0.0
        run_data['max_sim'] = np.max(all_sims) if all_sims else 0.0
        run_data['min_sim'] = np.min(all_sims) if all_sims else 0.0
        all_runs.append(run_data)

        print(f"本轮完成 | 耗时: {elapsed:.1f}s | 总体平均相似度: {run_data['overall_avg']:.4f}")

    print("\n生成最终 Excel 报告...")
    generate_excel_report(all_runs, similar_group, isolated_group, report_dir)
    print("\n全部 20 轮训练完成！")

if __name__ == "__main__":
    run_20_times_training()