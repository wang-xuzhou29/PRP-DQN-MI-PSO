import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

def execute_validation_rules(case):
        x, y, z = case.x, case.y, case.z
        triggered = set()

        if (x > 60 and x < 70 and y > 225 and y < 235) != (x > 40 and x < 70 and y > 225 and y < 235): triggered.add(1)
        if (x > 60 and x < 70 and y > 225 and y < 235) != (x > 20 and x < 70 and y > 225 and y < 235): triggered.add(2)
        if (x > 60 and x < 70 and y > 225 and y < 235) != (x > 60 and x < 70 and y > 220 and y < 235): triggered.add(3)
        if (x > 60 and x < 70 and y > 225 and y < 235) != (x > 60 and x < 70 and y > 225 and y < 230): triggered.add(4)

        # 第2组: 流量和电流的复合条件 (分支5-8)
        if (x > 55 and x < 75 and z > 11 and z < 14) != (x > 45 and x < 75 and z > 11 and z < 14): triggered.add(5)
        if (x > 55 and x < 75 and z > 11 and z < 14) != (x > 35 and x < 75 and z > 11 and z < 14): triggered.add(6)
        if (x > 55 and x < 75 and z > 11 and z < 14) != (x > 55 and x < 75 and z > 4 and z < 14): triggered.add(7)
        if (x > 55 and x < 75 and z > 11 and z < 14) != (x > 55 and x < 75 and z > 11 and z < 4): triggered.add(8)

        # 第3组: 电压和电流的复合条件 (分支9)
        if (y > 225 and y < 235 and z > 11 and z < 14) != (y > 220 and y < 235 and z > 11 and z < 14): triggered.add(9)

        # 第4组: 流量精确区间判断 (分支10-12)
        if (x > 55 and x < 67) != (x > 55 and x < 75): triggered.add(10)
        if (x > 55 and x < 67) != (x > 55 and x < 80): triggered.add(11)
        if (x > 55 and x < 67) != (x > 35 and x < 67): triggered.add(12)

        # 第5组: 电压和电流的精确条件 (分支13-18)
        if (y > 228 and y < 232) != (y > 228 and y < 230): triggered.add(13)

        if (abs(z - 12.5) < 0.8) != (abs(z - 11.5) < 0.8): triggered.add(14)
        if (abs(z - 12.5) < 0.8) != (abs(z - 10.5) < 0.8): triggered.add(15)
        if ((x - 50) / 30 > 0.45 and (x - 50) / 30 < 0.55) != (
                (x - 50) / 30 > 0.45 and (x - 50) / 20 < 0.55): triggered.add(16)
        if ((x - 50) / 30 > 0.45 and (x - 50) / 30 < 0.55) != (
                (x - 50) / 30 > 0.45 and (x - 50) / 30 < 0.85): triggered.add(17)
        if (abs(z - 12.5) < 0.8) != (abs(z - 13.5) < 0.8): triggered.add(18)

        # 第6组: 三参数复合条件 (分支19-22)
        if (x > 58 and x < 72 and y > 223 and y < 237 and z > 10.5 and z < 14.5) != (
                x > 68 and x < 72 and y > 223 and y < 237 and z > 10.5 and z < 14.5): triggered.add(19)
        if (x > 58 and x < 72 and y > 223 and y < 237 and z > 10.5 and z < 14.5) != (
                x > 48 and x < 72 and y > 223 and y < 237 and z > 10.5 and z < 14.5): triggered.add(20)
        if (x > 58 and x < 72 and y > 223 and y < 237 and z > 10.5 and z < 14.5) != (
                x > 58 and x < 72 and y > 233 and y < 237 and z > 10.5 and z < 14.5): triggered.add(21)
        if (x > 58 and x < 72 and y > 223 and y < 237 and z > 10.5 and z < 14.5) != (
                x > 58 and x < 72 and y > 223 and y < 227 and z > 10.5 and z < 14.5): triggered.add(22)

        # 第7组: 比值计算条件 (分支23-29)
        if y != 0:  # 避免除零错误
            if (x / (y / 10) > 2.5 and x / (y / 10) < 3.5) != (
                    x / (y / 13) > 2.5 and x / (y / 10) < 3.5): triggered.add(23)
            if (x / (y / 10) > 2.5 and x / (y / 10) < 3.5) != (x / (y / 6) > 2.5 and x / (y / 10) < 3.5): triggered.add(
                24)
            if (x / (y / 10) > 2.5 and x / (y / 10) < 3.5) != (
                    x / (y / 10) > 1.5 and x / (y / 10) < 3.5): triggered.add(25)
            if (x / (y / 10) > 2.5 and x / (y / 10) < 3.5) != (
                    x / (y / 10) > 2.5 and x / (y / 10) < 5.5): triggered.add(26)
            if (x / (y / 10) > 2.5 and x / (y / 10) < 3.5) != (
                    x / (y / 10) > 2.5 and x / ((y + 100) / 10) < 3.5): triggered.add(27)
            if (x / (y / 10) > 2.5 and x / (y / 10) < 3.5) != (
                    x / (y / 10) > 2.5 and x / (y / 11) < 3.5): triggered.add(28)
            if (x / (y / 10) > 2.5 and x / (y / 10) < 3.5) != (
                    x / (y / 10) > 2.5 and x / (y / 15) < 3.5): triggered.add(29)

        # 第8组: 流量电流比值条件 (分支30-35)
        if z != 0:  # 避免除零错误
            if (x / z > 4.5 and x / z < 6.5) != (x / z > 5.5 and x / z < 6.5): triggered.add(30)
            if (x / z > 4.5 and x / z < 6.5) != (x / z > 4.5 and x / z < 7.5): triggered.add(31)
            if (x / z > 4.5 and x / z < 6.5) != (x / z > 4.5 and x / z < 8.5): triggered.add(32)
            if (x / z > 4.5 and x / z < 6.5) != (y / z > 4.5 and x / z < 6.5): triggered.add(33)
            if y != 0:
                if (x / z > 4.5 and x / z < 6.5) != (x / (y / 10) > 4.5 and x / z < 6.5): triggered.add(34)
                if (x / z > 4.5 and x / z < 6.5) != (x / z > 4.5 and x / (y / 10) < 6.5): triggered.add(35)

        # 第9组: 乘积条件 (分支36-43)
        if (x * z > 600 and x * z < 1000) != (x * z > 700 and x * z < 1000): triggered.add(36)
        if (x * z > 600 and x * z < 1000) != (x * z > 800 and x * z < 1000): triggered.add(37)
        if (x * z > 600 and x * z < 1000) != (x * z > 900 and x * z < 1000): triggered.add(38)
        if (x * z > 600 and x * z < 1000) != (x * z > 600 and x * z < 1100): triggered.add(39)
        if (x * z > 600 and x * z < 1000) != (x * z > 600 and x * z < 1200): triggered.add(40)
        if (x * z > 600 and x * z < 1000) != (x * y > 600 and x * z < 1000): triggered.add(41)
        if (x * z > 600 and x * z < 1000) != (x * z > 600 and x * y < 1000): triggered.add(42)
        if (x * z > 600 and x * z < 1000) != (x * z > 600 and y * z < 1000): triggered.add(43)

        # 第10组: 加权平均条件 (分支44-50)
        if ((x + y / 10) / 2 > 40 and (x + y / 10) / 2 < 50) != (
                (x + z / 10) / 2 > 40 and (x + y / 10) / 2 < 50): triggered.add(44)
        if ((x + y / 10) / 2 > 40 and (x + y / 10) / 2 < 50) != (
                (z + y / 10) / 2 > 40 and (x + y / 10) / 2 < 50): triggered.add(45)
        if ((x + y / 10) / 2 > 40 and (x + y / 10) / 2 < 50) != (
                (x + y / 10) / 3 > 40 and (x + y / 10) / 2 < 50): triggered.add(46)
        if ((x + y / 10) / 2 > 40 and (x + y / 10) / 2 < 50) != (
                (x + y / 10) / 2 > 30 and (x + y / 10) / 2 < 50): triggered.add(47)
        if ((x + y / 10) / 2 > 40 and (x + y / 10) / 2 < 50) != (
                (x + y / 10) / 2 > 40 and (x + y / 14) / 2 < 50): triggered.add(48)
        if ((x + y / 10) / 2 > 40 and (x + y / 10) / 2 < 50) != (
                (x + y / 10) / 2 > 40 and (x + y / 10) / 3 < 50): triggered.add(49)
        if ((x + y / 10) / 2 > 40 and (x + y / 10) / 2 < 50) != (
                (x + y / 10) / 2 > 40 and (x + y / 10) / 2 < 60): triggered.add(50)

        # 第11组: 复杂乘积条件 (分支51-59)
        if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != (
                (x * (y / 10) * z) > 15000 and (x * (y / 10) * z) < 26000): triggered.add(51)
        if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != (
                (x * (y / 5) * z) > 18000 and (x * (y / 10) * z) < 26000): triggered.add(52)
        if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != (
                (x * (y / 10) * z) > 18000 and (x * (y / 5) * z) < 26000): triggered.add(53)
        if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != (
                (x * (y / 10) * z) > 18000 and (z * (y / 10) * z) < 26000): triggered.add(54)
        if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != (
                (z * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000): triggered.add(55)
        if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != (
                (x * z * z) > 18000 and (x * (y / 10) * z) < 26000): triggered.add(56)
        if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != (
                (x * x * z) > 18000 and (x * z * z) < 26000): triggered.add(57)
        if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != (
                (x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 20000): triggered.add(58)
        if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != (
                (x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 19000): triggered.add(59)

        # 第12组: 幂次条件 (分支60-67)
        if (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4500) != (
                x ** 2 + (y / 10) ** 3 + (z * 5) ** 2 > 4500): triggered.add(60)
        if (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4500) != (
                x ** 3 + (y / 10) ** 2 + (z * 5) ** 2 > 4500): triggered.add(61)
        if (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4500) != (
                x ** 2 + (y / 10) ** 2 + (z * 5) ** 3 > 4500): triggered.add(62)
        if (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4500) != (
                x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 3500): triggered.add(63)
        if (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4500) != (
                x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 5500): triggered.add(64)
        if (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4500) != (
                x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4000): triggered.add(65)
        if (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4500) != (
                x ** 2 + (y / 10) ** 2 + (z * 10) ** 2 > 4500): triggered.add(66)
        if (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4500) != (
                x ** 2 + (y / 5) ** 2 + (z * 5) ** 2 > 4500): triggered.add(67)

        # 第13组: 归一化条件 (分支68-77)
        if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 65 > 0.95 and x / 65 < 1.15): triggered.add(68)
        if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 55 > 0.85 and x / 65 < 1.15): triggered.add(69)
        if (x / 65 > 0.85 and x / 65 < 1.15) != (y / 65 > 0.85 and x / 65 < 1.15): triggered.add(70)
        if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 65 > 0.85 and y / 65 < 1.15): triggered.add(71)
        if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 65 > 0.85 and x / 45 < 1.15): triggered.add(72)
        if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 65 > 0.85 and x / 55 < 1.15): triggered.add(73)
        if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 65 > 0.85 and x / 75 < 1.15): triggered.add(74)
        if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 65 > 0.85 and x / 65 < 2.15): triggered.add(75)
        if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 65 > 0.85 and x / 65 < 1.0): triggered.add(76)
        if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 65 > 0.85 and x / 65 < 1.05): triggered.add(77)

        # 第14组: 阈值条件 (分支78-84)
        if (x < 45 and y < 220) != (x < 55 and y < 220): triggered.add(78)
        if (x < 45 and y < 220) != (x < 35 and y < 220): triggered.add(79)
        if (x < 45 and y < 220) != (x < 25 and y < 220): triggered.add(80)
        if (x < 45 and y < 220) != (x < 45 and y < 230): triggered.add(81)
        if (x < 45 and y < 220) != (x < 45 and y < 240): triggered.add(82)
        if (x < 45 and y < 220) != (x < 45 and y < 210): triggered.add(83)
        if (x < 45 and y < 220) != (x < 45 and y > 220): triggered.add(84)

        # 第15组: 高阈值条件 (分支85-90)
        if (y > 240 and z > 16) != (y > 230 and z > 16): triggered.add(85)
        if (y > 240 and z > 16) != (y > 220 and z > 16): triggered.add(86)
        if (y > 240 and z > 16) != (y > 210 and z > 16): triggered.add(87)
        if (y > 240 and z > 16) != (y > 240 and x > 16): triggered.add(88)
        if (y > 240 and z > 16) != (y > 240 and z < 16): triggered.add(89)
        if (y > 240 and z > 16) != (y < 240 and z > 16): triggered.add(90)

        # 第16组: 复合逻辑条件 (分支91-96)
        if ((x + y / 10 + z * 5) < 95 or (x + y / 10 + z * 5) > 135) != (
                (x + z + z * 5) < 95 or (x + y / 10 + z * 5) > 135): triggered.add(91)
        if ((x + y / 10 + z * 5) < 95 or (x + y / 10 + z * 5) > 135) != (
                (z + y / 10 + z * 5) < 95 or (x + y / 10 + z * 5) > 135): triggered.add(92)
        if ((x + y / 10 + z * 5) < 95 or (x + y / 10 + z * 5) > 135) != (
                (x + y / 10 + z * 5) < 95 or (x + z + z * 5) > 135): triggered.add(93)
        if ((x + y / 10 + z * 5) < 95 or (x + y / 10 + z * 5) > 135) != (
                (x + y / 10 + z * 5) < 95 or (z + y / 10 + z * 5) > 135): triggered.add(94)
        if ((x + y / 10 + z * 5) < 95 or (x + y / 10 + z * 5) > 135) != (
                (x + y / 10 + z * 5) < 95 or (x + y / 10 + x * 5) > 135): triggered.add(95)
        if ((x + y / 10 + z * 5) < 95 or (x + y / 10 + z * 5) > 135) != (
                (x + y / 10 + z * 5) < 95 and (x + y / 10 + z * 5) > 135): triggered.add(96)

        # 第17组: 精确区间条件 (分支97-100)
        if (x > 64 and x < 66 and y > 229 and y < 231) != (x > 60 and x < 66 and y > 229 and y < 231): triggered.add(97)
        if (x > 64 and x < 66 and y > 229 and y < 231) != (x > 54 and x < 66 and y > 229 and y < 231): triggered.add(98)
        if (x > 64 and x < 66 and y > 229 and y < 231) != (x > 64 and x < 76 and y > 229 and y < 231): triggered.add(99)
        if (x > 64 and x < 66 and y > 229 and y < 231) != (x > 64 and x < 66 and y < 229 and y < 231): triggered.add(
            100)

        # 第18组: 标准化条件 (分支101-106)
        if ((x - 50) / 30 > 0.45 and (x - 50) / 30 < 0.55) != (
                (x - 40) / 30 > 0.45 and (x - 50) / 30 < 0.55): triggered.add(101)
        if ((x - 50) / 30 > 0.45 and (x - 50) / 30 < 0.55) != (
                (x - 50) / 20 > 0.45 and (x - 50) / 30 < 0.55): triggered.add(102)
        if ((x - 50) / 30 > 0.45 and (x - 50) / 30 < 0.55) != (
                (x - 50) / 30 > 0.65 and (x - 50) / 30 < 0.55): triggered.add(103)
        if ((x - 50) / 30 > 0.45 and (x - 50) / 30 < 0.55) != (
                (x - 50) / 30 > 0.45 and (x - 40) / 30 < 0.55): triggered.add(104)


        return triggered

def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """"""
    generated_path = execute_validation_rules(particle[0], particle[1], particle[2])

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # : x:1-50, y:1-50, z:1-50
        self.bounds = bounds if bounds else [(20, 80), (200, 250), (4, 20)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """target pathsPSO"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(particles[i][0], particles[i][1], particles[i][2]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(gbest_particle[0], gbest_particle[1], gbest_particle[2]),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """Path PSO"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - Path ")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations")
    print(f"Path : {len(target_paths)}")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"Path {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "" if result['success'] else f"({result['best_fitness']:.3f})"
        print(f"{status} | {result['time']:.2f}s | iterations{result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f": {success_count}/{len(target_paths)} ({success_rate:.1f}%) | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs}")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations, {len(target_paths)}Path ")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"---  {run_idx}/{num_runs} ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f": {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs} runcompleted | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 1: 
    ws1 = wb.active
    ws1.title = ""
    ws1.sheet_view.showGridLines = False

    headers = ["Run", "", "", "", "Average Iterations", "(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f" {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # 2: Path 
    ws2 = wb.create_sheet(title="Path ")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Path ID", "", "", "", "Average Iterations", "Minimum Iterations", "Maximum Iterations"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"Path  {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # 3: 
    ws3 = wb.create_sheet(title="")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Path ", "", "(x,y,z)", "", "Iterations", "Path "]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"{run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # 4: target paths
    ws4 = wb.create_sheet(title="target paths")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Path ID", "target paths", ""]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"Path  {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f" : {filename}")
    print(f"{'=' * 70}")
    print(f":")
    print(f"  1.        - {len(all_results)} run")
    print(f"  2. Path        - Path ")
    print(f"  3.    -  runPath ")
    print(f"  4. target paths       - target paths")
    print(f"{'=' * 70}\n")

    return filename


def main():


    target_paths = [
        {4, 8, 13, 15, 18, 19, 21, 22, 24, 29, 30, 34, 37, 38, 42, 43, 44, 45, 46, 51, 52, 57, 68, 71, 72, 93, 94, 96,
         97,
         98, 101, 102},
        {1, 2, 8, 13, 15, 18, 19, 21, 22, 24, 29, 30, 34, 37, 38, 42, 43, 44, 45, 46, 51, 52, 57, 68, 71, 72, 93, 94,
         96,
         98, 101, 102},
        {4, 8, 10, 11, 13, 14, 15, 17, 19, 21, 22, 24, 29, 30, 34, 38, 42, 43, 44, 45, 46, 53, 55, 56, 58, 59, 71, 72,
         73,
         76, 94, 96},
        {4, 7, 10, 11, 13, 14, 15, 17, 21, 22, 24, 29, 34, 37, 38, 42, 43, 44, 45, 46, 51, 52, 57, 71, 72, 73, 76, 77,
         93,
         94, 96, 99},
        {4, 8, 13, 14, 15, 16, 19, 21, 22, 24, 29, 30, 34, 38, 42, 43, 44, 45, 46, 53, 55, 56, 59, 71, 72, 73, 76, 94,
         96,
         103, 104},
        {4, 7, 13, 15, 16, 24, 29, 31, 32, 35, 36, 37, 38, 42, 43, 44, 45, 46, 51, 52, 57, 71, 72, 73, 76, 93, 94, 96,
         99,
         103, 104},
        {7, 10, 11, 14, 15, 17, 24, 29, 31, 32, 35, 37, 38, 42, 43, 44, 45, 46, 53, 55, 56, 59, 71, 72, 73, 76, 77, 88,
         89,
         94, 96},
        {3, 8, 9, 10, 11, 14, 15, 17, 21, 24, 29, 30, 34, 38, 42, 43, 44, 45, 46, 53, 55, 56, 58, 59, 71, 72, 73, 76,
         77,
         94, 96},
        {3, 8, 9, 14, 15, 16, 19, 21, 24, 29, 30, 34, 38, 42, 43, 44, 45, 46, 53, 55, 56, 71, 72, 73, 76, 94, 96, 100,
         103,
         104},
        {7, 10, 11, 13, 15, 17, 24, 28, 29, 31, 32, 35, 37, 38, 42, 43, 44, 45, 46, 51, 52, 57, 71, 72, 73, 76, 77, 93,
         94,
         96},
        {7, 15, 16, 24, 28, 29, 31, 32, 35, 36, 37, 38, 42, 43, 44, 45, 46, 52, 57, 71, 72, 73, 76, 92, 95, 100, 103,
         104},
        {1, 2, 13, 18, 19, 21, 22, 24, 29, 33, 38, 42, 43, 44, 45, 46, 53, 55, 56, 59, 68, 71, 72, 94, 96, 98, 101,
         102},
        {1, 2, 8, 15, 18, 20, 23, 25, 30, 34, 36, 37, 38, 42, 43, 44, 45, 46, 51, 52, 57, 68, 71, 72, 93, 94, 96, 101},
        {1, 2, 5, 6, 12, 13, 15, 18, 20, 23, 25, 30, 34, 36, 37, 38, 42, 43, 47, 51, 52, 57, 69, 70, 93, 94, 96, 101},
        {4, 13, 24, 29, 33, 39, 40, 44, 45, 46, 53, 55, 56, 58, 59, 68, 71, 72, 86, 87, 90, 94, 96, 97, 98, 101, 102},
        {1, 2, 24, 29, 33, 39, 40, 44, 45, 46, 53, 55, 56, 58, 59, 68, 71, 72, 85, 86, 87, 90, 94, 96, 101, 102},
        {8, 10, 11, 14, 15, 17, 26, 27, 34, 42, 43, 44, 45, 46, 53, 55, 56, 58, 59, 71, 72, 73, 76, 77, 94, 96},
        {11, 13, 17, 24, 28, 29, 32, 35, 36, 37, 38, 42, 43, 44, 45, 46, 51, 52, 57, 74, 75, 93, 94, 96, 99},
        {11, 15, 18, 24, 28, 29, 31, 32, 35, 42, 43, 48, 49, 50, 53, 55, 56, 58, 59, 74, 75, 88, 89, 94, 96},
        {4, 13, 16, 24, 29, 33, 40, 44, 45, 46, 54, 57, 71, 72, 73, 86, 87, 90, 94, 96, 97, 98, 103, 104},
        {1, 2, 13, 24, 29, 35, 41, 44, 45, 46, 60, 61, 62, 63, 65, 66, 67, 68, 71, 72, 91, 92, 95, 101},
        {4, 7, 13, 16, 24, 29, 35, 41, 44, 45, 46, 57, 64, 71, 72, 73, 91, 92, 95, 97, 98, 103, 104},
        {1, 2, 12, 25, 33, 36, 37, 38, 42, 43, 47, 51, 52, 57, 70, 82, 84, 85, 86, 87, 90, 94, 96},
        {1, 2, 12, 23, 25, 33, 38, 42, 43, 47, 53, 55, 56, 59, 70, 81, 82, 84, 86, 87, 90, 96},
        {12, 15, 25, 33, 41, 47, 60, 61, 62, 63, 65, 66, 67, 70, 79, 80, 83, 84, 92, 95},
        {12, 23, 25, 32, 35, 41, 47, 57, 60, 61, 62, 63, 65, 66, 67, 69, 70, 78, 92, 95},
    ]

    print("=" * 70)
    print("baseline PSO")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("Program completed")


if __name__ == "__main__":
    main()