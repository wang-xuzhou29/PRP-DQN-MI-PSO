import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
import numpy as np
import random
import time
import math
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
from datetime import datetime

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")


# === 目标路径组 ===
targetPaths = [
    {4, 8, 13, 15, 18, 19, 21, 22, 24, 29, 30, 34, 37, 38, 42, 43, 44, 45, 46, 51, 52, 57, 68, 71, 72, 93, 94, 96, 97,
     98, 101, 102},
    {1, 2, 8, 13, 15, 18, 19, 21, 22, 24, 29, 30, 34, 37, 38, 42, 43, 44, 45, 46, 51, 52, 57, 68, 71, 72, 93, 94, 96,
     98, 101, 102},
    {4, 8, 10, 11, 13, 14, 15, 17, 19, 21, 22, 24, 29, 30, 34, 38, 42, 43, 44, 45, 46, 53, 55, 56, 58, 59, 71, 72, 73,
     76, 94, 96},
    {4, 7, 10, 11, 13, 14, 15, 17, 21, 22, 24, 29, 34, 37, 38, 42, 43, 44, 45, 46, 51, 52, 57, 71, 72, 73, 76, 77, 93,
     94, 96, 99},
    {4, 8, 13, 14, 15, 16, 19, 21, 22, 24, 29, 30, 34, 38, 42, 43, 44, 45, 46, 53, 55, 56, 59, 71, 72, 73, 76, 94, 96,
     103, 104},
    {4, 7, 13, 15, 16, 24, 29, 31, 32, 35, 36, 37, 38, 42, 43, 44, 45, 46, 51, 52, 57, 71, 72, 73, 76, 93, 94, 96, 99,
     103, 104},
    {7, 10, 11, 14, 15, 17, 24, 29, 31, 32, 35, 37, 38, 42, 43, 44, 45, 46, 53, 55, 56, 59, 71, 72, 73, 76, 77, 88, 89,
     94, 96},
    {3, 8, 9, 10, 11, 14, 15, 17, 21, 24, 29, 30, 34, 38, 42, 43, 44, 45, 46, 53, 55, 56, 58, 59, 71, 72, 73, 76, 77,
     94, 96},
    {3, 8, 9, 14, 15, 16, 19, 21, 24, 29, 30, 34, 38, 42, 43, 44, 45, 46, 53, 55, 56, 71, 72, 73, 76, 94, 96, 100, 103,
     104},
    {7, 10, 11, 13, 15, 17, 24, 28, 29, 31, 32, 35, 37, 38, 42, 43, 44, 45, 46, 51, 52, 57, 71, 72, 73, 76, 77, 93, 94,
     96},
    {7, 15, 16, 24, 28, 29, 31, 32, 35, 36, 37, 38, 42, 43, 44, 45, 46, 52, 57, 71, 72, 73, 76, 92, 95, 100, 103, 104},
    {1, 2, 13, 18, 19, 21, 22, 24, 29, 33, 38, 42, 43, 44, 45, 46, 53, 55, 56, 59, 68, 71, 72, 94, 96, 98, 101, 102},
    {1, 2, 8, 15, 18, 20, 23, 25, 30, 34, 36, 37, 38, 42, 43, 44, 45, 46, 51, 52, 57, 68, 71, 72, 93, 94, 96, 101},
    {1, 2, 5, 6, 12, 13, 15, 18, 20, 23, 25, 30, 34, 36, 37, 38, 42, 43, 47, 51, 52, 57, 69, 70, 93, 94, 96, 101},
    {4, 13, 24, 29, 33, 39, 40, 44, 45, 46, 53, 55, 56, 58, 59, 68, 71, 72, 86, 87, 90, 94, 96, 97, 98, 101, 102},
    {1, 2, 24, 29, 33, 39, 40, 44, 45, 46, 53, 55, 56, 58, 59, 68, 71, 72, 85, 86, 87, 90, 94, 96, 101, 102},
    {8, 10, 11, 14, 15, 17, 26, 27, 34, 42, 43, 44, 45, 46, 53, 55, 56, 58, 59, 71, 72, 73, 76, 77, 94, 96},
    {11, 13, 17, 24, 28, 29, 32, 35, 36, 37, 38, 42, 43, 44, 45, 46, 51, 52, 57, 74, 75, 93, 94, 96, 99},
    {11, 15, 18, 24, 28, 29, 31, 32, 35, 42, 43, 48, 49, 50, 53, 55, 56, 58, 59, 74, 75, 88, 89, 94, 96},
    {4, 13, 16, 24, 29, 33, 40, 44, 45, 46, 54, 57, 71, 72, 73, 86, 87, 90, 94, 96, 97, 98, 103, 104},
    {1, 2, 13, 24, 29, 35, 41, 44, 45, 46, 60, 61, 62, 63, 65, 66, 67, 68, 71, 72, 91, 92, 95, 101},
    {4, 7, 13, 16, 24, 29, 35, 41, 44, 45, 46, 57, 64, 71, 72, 73, 91, 92, 95, 97, 98, 103, 104},
    {1, 2, 12, 25, 33, 36, 37, 38, 42, 43, 47, 51, 52, 57, 70, 82, 84, 85, 86, 87, 90, 94, 96},
    {1, 2, 12, 23, 25, 33, 38, 42, 43, 47, 53, 55, 56, 59, 70, 81, 82, 84, 86, 87, 90, 96},
    {12, 15, 25, 33, 41, 47, 60, 61, 62, 63, 65, 66, 67, 70, 79, 80, 83, 84, 92, 95},
    {12, 23, 25, 32, 35, 41, 47, 57, 60, 61, 62, 63, 65, 66, 67, 69, 70, 78, 92, 95},

]

# === 配置（取值范围：x:20~80 流量, y:200~250 电压, z:4~20 电流） ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'ACTION_DIM': 3,
    'MIN_VALUES': np.array([20, 200, 4]),
    'MAX_VALUES': np.array([80, 250, 20]),
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 3,
    'PPO_BATCH_SIZE': 64,
    'PPO_EPOCHS': 4,
    'CLIP_EPSILON': 0.2,
    'GAMMA': 0.99,
    'GAE_LAMBDA': 0.95,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'ACTOR_LR': 3e-4,
    'CRITIC_LR': 3e-4,
    'NUM_RUNS': 20,
    'TOP_K_SAMPLES': 20,
    'TARGET_PATHS': targetPaths
}


def section3_thermal_electrical_flow_hybrid(x, y, z):
    triggered = set()

    # 第1组: 流量电压复合基准条件 (分支1-4)
    if (x > 60 and x < 70 and y > 225 and y < 235) != (x > 40 and x < 70 and y > 225 and y < 235): triggered.add(1)
    if (x > 60 and x < 70 and y > 225 and y < 235) != (x > 20 and x < 70 and y > 225 and y < 235): triggered.add(2)
    if (x > 60 and x < 70 and y > 225 and y < 235) != (x > 60 and x < 70 and y > 220 and y < 235): triggered.add(3)
    if (x > 60 and x < 70 and y > 225 and y < 235) != (x > 60 and x < 70 and y > 225 and y < 230): triggered.add(4)

    # 第2组: 流量和电流的复合条件 (分支5-8)
    if (x > 55 and x < 75 and z > 11 and z < 14) != (x > 45 and x < 75 and z > 11 and z < 14): triggered.add(5)
    if (x > 55 and x < 75 and z > 11 and z < 14) != (x > 35 and x < 75 and z > 11 and z < 14): triggered.add(6)
    if (x > 55 and x < 75 and z > 11 and z < 14) != (x > 55 and x < 75 and z > 4 and z < 14): triggered.add(7)
    if (x > 55 and x < 75 and z > 11 and z < 14) != (x > 55 and x < 75 and z > 11 and z < 4): triggered.add(8)

    # 第3组: 电压和电流的复合条件 (分支9)
    if (y > 225 and y < 235 and z > 11 and z < 14) != (y > 220 and y < 235 and z > 11 and z < 14): triggered.add(9)

    # 第4组: 流量精确区间判断 (分支10-12)
    if (x > 55 and x < 67) != (x > 55 and x < 75): triggered.add(10)
    if (x > 55 and x < 67) != (x > 55 and x < 80): triggered.add(11)
    if (x > 55 and x < 67) != (x > 35 and x < 67): triggered.add(12)

    # 第5组: 电压和电流的精确条件 (分支13-18)
    if (y > 228 and y < 232) != (y > 228 and y < 230): triggered.add(13)
    if (abs(z - 12.5) < 0.8) != (abs(z - 11.5) < 0.8): triggered.add(14)
    if (abs(z - 12.5) < 0.8) != (abs(z - 10.5) < 0.8): triggered.add(15)
    if ((x - 50) / 30 > 0.45 and (x - 50) / 30 < 0.55) != (
            (x - 50) / 30 > 0.45 and (x - 50) / 20 < 0.55): triggered.add(16)
    if ((x - 50) / 30 > 0.45 and (x - 50) / 30 < 0.55) != (
            (x - 50) / 30 > 0.45 and (x - 50) / 30 < 0.85): triggered.add(17)
    if (abs(z - 12.5) < 0.8) != (abs(z - 13.5) < 0.8): triggered.add(18)

    # 第6组: 三参数复合条件 (分支19-22)
    if (x > 58 and x < 72 and y > 223 and y < 237 and z > 10.5 and z < 14.5) != (
            x > 68 and x < 72 and y > 223 and y < 237 and z > 10.5 and z < 14.5): triggered.add(19)
    if (x > 58 and x < 72 and y > 223 and y < 237 and z > 10.5 and z < 14.5) != (
            x > 48 and x < 72 and y > 223 and y < 237 and z > 10.5 and z < 14.5): triggered.add(20)
    if (x > 58 and x < 72 and y > 223 and y < 237 and z > 10.5 and z < 14.5) != (
            x > 58 and x < 72 and y > 233 and y < 237 and z > 10.5 and z < 14.5): triggered.add(21)
    if (x > 58 and x < 72 and y > 223 and y < 237 and z > 10.5 and z < 14.5) != (
            x > 58 and x < 72 and y > 223 and y < 227 and z > 10.5 and z < 14.5): triggered.add(22)

    # 第7组: 比值计算条件 (分支23-29)
    if y != 0:  # 避免除零错误
        if (x / (y / 10) > 2.5 and x / (y / 10) < 3.5) != (
                x / (y / 13) > 2.5 and x / (y / 10) < 3.5): triggered.add(23)
        if (x / (y / 10) > 2.5 and x / (y / 10) < 3.5) != (x / (y / 6) > 2.5 and x / (y / 10) < 3.5): triggered.add(24)
        if (x / (y / 10) > 2.5 and x / (y / 10) < 3.5) != (
                x / (y / 10) > 1.5 and x / (y / 10) < 3.5): triggered.add(25)
        if (x / (y / 10) > 2.5 and x / (y / 10) < 3.5) != (
                x / (y / 10) > 2.5 and x / (y / 10) < 5.5): triggered.add(26)
        if (x / (y / 10) > 2.5 and x / (y / 10) < 3.5) != (
                x / (y / 10) > 2.5 and x / ((y + 100) / 10) < 3.5): triggered.add(27)
        if (x / (y / 10) > 2.5 and x / (y / 10) < 3.5) != (
                x / (y / 10) > 2.5 and x / (y / 11) < 3.5): triggered.add(28)
        if (x / (y / 10) > 2.5 and x / (y / 10) < 3.5) != (
                x / (y / 10) > 2.5 and x / (y / 15) < 3.5): triggered.add(29)

    # 第8组: 流量电流比值条件 (分支30-35)
    if z != 0:  # 避免除零错误
        if (x / z > 4.5 and x / z < 6.5) != (x / z > 5.5 and x / z < 6.5): triggered.add(30)
        if (x / z > 4.5 and x / z < 6.5) != (x / z > 4.5 and x / z < 7.5): triggered.add(31)
        if (x / z > 4.5 and x / z < 6.5) != (x / z > 4.5 and x / z < 8.5): triggered.add(32)
        if (x / z > 4.5 and x / z < 6.5) != (y / z > 4.5 and x / z < 6.5): triggered.add(33)
        if y != 0:
            if (x / z > 4.5 and x / z < 6.5) != (x / (y / 10) > 4.5 and x / z < 6.5): triggered.add(34)
            if (x / z > 4.5 and x / z < 6.5) != (x / z > 4.5 and x / (y / 10) < 6.5): triggered.add(35)

    # 第9组: 乘积条件 (分支36-43)
    if (x * z > 600 and x * z < 1000) != (x * z > 700 and x * z < 1000): triggered.add(36)
    if (x * z > 600 and x * z < 1000) != (x * z > 800 and x * z < 1000): triggered.add(37)
    if (x * z > 600 and x * z < 1000) != (x * z > 900 and x * z < 1000): triggered.add(38)
    if (x * z > 600 and x * z < 1000) != (x * z > 600 and x * z < 1100): triggered.add(39)
    if (x * z > 600 and x * z < 1000) != (x * z > 600 and x * z < 1200): triggered.add(40)
    if (x * z > 600 and x * z < 1000) != (x * y > 600 and x * z < 1000): triggered.add(41)
    if (x * z > 600 and x * z < 1000) != (x * z > 600 and x * y < 1000): triggered.add(42)
    if (x * z > 600 and x * z < 1000) != (x * z > 600 and y * z < 1000): triggered.add(43)

    # 第10组: 加权平均条件 (分支44-50)
    if ((x + y / 10) / 2 > 40 and (x + y / 10) / 2 < 50) != (
            (x + z / 10) / 2 > 40 and (x + y / 10) / 2 < 50): triggered.add(44)
    if ((x + y / 10) / 2 > 40 and (x + y / 10) / 2 < 50) != (
            (z + y / 10) / 2 > 40 and (x + y / 10) / 2 < 50): triggered.add(45)
    if ((x + y / 10) / 2 > 40 and (x + y / 10) / 2 < 50) != (
            (x + y / 10) / 3 > 40 and (x + y / 10) / 2 < 50): triggered.add(46)
    if ((x + y / 10) / 2 > 40 and (x + y / 10) / 2 < 50) != (
            (x + y / 10) / 2 > 30 and (x + y / 10) / 2 < 50): triggered.add(47)
    if ((x + y / 10) / 2 > 40 and (x + y / 10) / 2 < 50) != (
            (x + y / 10) / 2 > 40 and (x + y / 14) / 2 < 50): triggered.add(48)
    if ((x + y / 10) / 2 > 40 and (x + y / 10) / 2 < 50) != (
            (x + y / 10) / 2 > 40 and (x + y / 10) / 3 < 50): triggered.add(49)
    if ((x + y / 10) / 2 > 40 and (x + y / 10) / 2 < 50) != (
            (x + y / 10) / 2 > 40 and (x + y / 10) / 2 < 60): triggered.add(50)

    # 第11组: 复杂乘积条件 (分支51-59)
    if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != (
            (x * (y / 10) * z) > 15000 and (x * (y / 10) * z) < 26000): triggered.add(51)
    if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != (
            (x * (y / 5) * z) > 18000 and (x * (y / 10) * z) < 26000): triggered.add(52)
    if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != (
            (x * (y / 10) * z) > 18000 and (x * (y / 5) * z) < 26000): triggered.add(53)
    if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != (
            (x * (y / 10) * z) > 18000 and (z * (y / 10) * z) < 26000): triggered.add(54)
    if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != (
            (z * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000): triggered.add(55)
    if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != (
            (x * z * z) > 18000 and (x * (y / 10) * z) < 26000): triggered.add(56)
    if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != (
            (x * x * z) > 18000 and (x * z * z) < 26000): triggered.add(57)
    if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != (
            (x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 20000): triggered.add(58)
    if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != (
            (x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 19000): triggered.add(59)

    # 第12组: 幂次条件 (分支60-67)
    if (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4500) != (
            x ** 2 + (y / 10) ** 3 + (z * 5) ** 2 > 4500): triggered.add(60)
    if (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4500) != (
            x ** 3 + (y / 10) ** 2 + (z * 5) ** 2 > 4500): triggered.add(61)
    if (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4500) != (
            x ** 2 + (y / 10) ** 2 + (z * 5) ** 3 > 4500): triggered.add(62)
    if (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4500) != (
            x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 3500): triggered.add(63)
    if (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4500) != (
            x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 5500): triggered.add(64)
    if (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4500) != (
            x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4000): triggered.add(65)
    if (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4500) != (
            x ** 2 + (y / 10) ** 2 + (z * 10) ** 2 > 4500): triggered.add(66)
    if (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4500) != (
            x ** 2 + (y / 5) ** 2 + (z * 5) ** 2 > 4500): triggered.add(67)

    # 第13组: 归一化条件 (分支68-77)
    if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 65 > 0.95 and x / 65 < 1.15): triggered.add(68)
    if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 55 > 0.85 and x / 65 < 1.15): triggered.add(69)
    if (x / 65 > 0.85 and x / 65 < 1.15) != (y / 65 > 0.85 and x / 65 < 1.15): triggered.add(70)
    if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 65 > 0.85 and y / 65 < 1.15): triggered.add(71)
    if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 65 > 0.85 and x / 45 < 1.15): triggered.add(72)
    if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 65 > 0.85 and x / 55 < 1.15): triggered.add(73)
    if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 65 > 0.85 and x / 75 < 1.15): triggered.add(74)
    if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 65 > 0.85 and x / 65 < 2.15): triggered.add(75)
    if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 65 > 0.85 and x / 65 < 1.0): triggered.add(76)
    if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 65 > 0.85 and x / 65 < 1.05): triggered.add(77)

    # 第14组: 阈值条件 (分支78-84)
    if (x < 45 and y < 220) != (x < 55 and y < 220): triggered.add(78)
    if (x < 45 and y < 220) != (x < 35 and y < 220): triggered.add(79)
    if (x < 45 and y < 220) != (x < 25 and y < 220): triggered.add(80)
    if (x < 45 and y < 220) != (x < 45 and y < 230): triggered.add(81)
    if (x < 45 and y < 220) != (x < 45 and y < 240): triggered.add(82)
    if (x < 45 and y < 220) != (x < 45 and y < 210): triggered.add(83)
    if (x < 45 and y < 220) != (x < 45 and y > 220): triggered.add(84)

    # 第15组: 高阈值条件 (分支85-90)
    if (y > 240 and z > 16) != (y > 230 and z > 16): triggered.add(85)
    if (y > 240 and z > 16) != (y > 220 and z > 16): triggered.add(86)
    if (y > 240 and z > 16) != (y > 210 and z > 16): triggered.add(87)
    if (y > 240 and z > 16) != (y > 240 and x > 16): triggered.add(88)
    if (y > 240 and z > 16) != (y > 240 and z < 16): triggered.add(89)
    if (y > 240 and z > 16) != (y < 240 and z > 16): triggered.add(90)

    # 第16组: 复合逻辑条件 (分支91-96)
    if ((x + y / 10 + z * 5) < 95 or (x + y / 10 + z * 5) > 135) != (
            (x + z + z * 5) < 95 or (x + y / 10 + z * 5) > 135): triggered.add(91)
    if ((x + y / 10 + z * 5) < 95 or (x + y / 10 + z * 5) > 135) != (
            (z + y / 10 + z * 5) < 95 or (x + y / 10 + z * 5) > 135): triggered.add(92)
    if ((x + y / 10 + z * 5) < 95 or (x + y / 10 + z * 5) > 135) != (
            (x + y / 10 + z * 5) < 95 or (x + z + z * 5) > 135): triggered.add(93)
    if ((x + y / 10 + z * 5) < 95 or (x + y / 10 + z * 5) > 135) != (
            (x + y / 10 + z * 5) < 95 or (z + y / 10 + z * 5) > 135): triggered.add(94)
    if ((x + y / 10 + z * 5) < 95 or (x + y / 10 + z * 5) > 135) != (
            (x + y / 10 + z * 5) < 95 or (x + y / 10 + x * 5) > 135): triggered.add(95)
    if ((x + y / 10 + z * 5) < 95 or (x + y / 10 + z * 5) > 135) != (
            (x + y / 10 + z * 5) < 95 and (x + y / 10 + z * 5) > 135): triggered.add(96)

    # 第17组: 精确区间条件 (分支97-100)
    if (x > 64 and x < 66 and y > 229 and y < 231) != (x > 60 and x < 66 and y > 229 and y < 231): triggered.add(97)
    if (x > 64 and x < 66 and y > 229 and y < 231) != (x > 54 and x < 66 and y > 229 and y < 231): triggered.add(98)
    if (x > 64 and x < 66 and y > 229 and y < 231) != (x > 64 and x < 76 and y > 229 and y < 231): triggered.add(99)
    if (x > 64 and x < 66 and y > 229 and y < 231) != (x > 64 and x < 66 and y < 229 and y < 231): triggered.add(100)

    # 第18组: 标准化条件 (分支101-104)
    if ((x - 50) / 30 > 0.45 and (x - 50) / 30 < 0.55) != (
            (x - 40) / 30 > 0.45 and (x - 50) / 30 < 0.55): triggered.add(101)
    if ((x - 50) / 30 > 0.45 and (x - 50) / 30 < 0.55) != (
            (x - 50) / 20 > 0.45 and (x - 50) / 30 < 0.55): triggered.add(102)
    if ((x - 50) / 30 > 0.45 and (x - 50) / 30 < 0.55) != (
            (x - 50) / 30 > 0.65 and (x - 50) / 30 < 0.55): triggered.add(103)
    if ((x - 50) / 30 > 0.45 and (x - 50) / 30 < 0.55) != (
            (x - 50) / 30 > 0.45 and (x - 40) / 30 < 0.55): triggered.add(104)

    return triggered


# === 执行函数绑定（修正为当前规则集） ===
execute_Tr = section3_thermal_electrical_flow_hybrid

# === 状态处理辅助函数 ===
def clip_state(state):
    return np.clip(state, EXPERIMENT_CONFIG['MIN_VALUES'], EXPERIMENT_CONFIG['MAX_VALUES'])

def denormalize_state(normalized_state):
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return normalized_state * (max_vals - min_vals) / 2 + (min_vals + max_vals) / 2

def coverage_similarity(triggered, target_path):
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0
    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)

def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']
    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']
    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']
    return reward

# === PPO Actor ===
class PPOActor(nn.Module):
    def __init__(self, state_dim=3, action_dim=3, hidden_dim=256):
        super(PPOActor, self).__init__()
        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)
        self.mean_head = nn.Linear(hidden_dim, action_dim)
        self.log_std_head = nn.Linear(hidden_dim, action_dim)
        self.action_scale = 10.0

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))
        x = F.relu(self.fc3(x))
        mean = torch.tanh(self.mean_head(x)) * self.action_scale
        log_std = torch.clamp(self.log_std_head(x), -20, 2)
        return mean, log_std

    def get_action_and_log_prob(self, state):
        mean, log_std = self.forward(state)
        std = torch.exp(log_std)
        dist = torch.distributions.Normal(mean, std)
        action = dist.sample()
        log_prob = dist.log_prob(action).sum(dim=-1)
        return action, log_prob

    def get_log_prob(self, state, action):
        mean, log_std = self.forward(state)
        std = torch.exp(log_std)
        dist = torch.distributions.Normal(mean, std)
        log_prob = dist.log_prob(action).sum(dim=-1)
        return log_prob

# === PPO Critic ===
class PPOCritic(nn.Module):
    def __init__(self, state_dim=3, hidden_dim=256):
        super(PPOCritic, self).__init__()
        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)
        self.value_head = nn.Linear(hidden_dim, 1)

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))
        x = F.relu(self.fc3(x))
        value = self.value_head(x)
        return value.squeeze(-1)

# === PPO Buffer ===
class PPOBuffer:
    def __init__(self):
        self.states = []
        self.actions = []
        self.rewards = []
        self.values = []
        self.log_probs = []
        self.dones = []
        self.path_indices = []
        self.similarities = []

    def store(self, state, action, reward, value, log_prob, done, path_idx, similarity):
        self.states.append(state)
        self.actions.append(action)
        self.rewards.append(reward)
        self.values.append(value)
        self.log_probs.append(log_prob)
        self.dones.append(done)
        self.path_indices.append(path_idx)
        self.similarities.append(similarity)

    def compute_advantages(self):
        rewards = np.array(self.rewards)
        values = np.array(self.values)
        dones = np.array(self.dones)
        advantages = np.zeros_like(rewards)
        returns = np.zeros_like(rewards)
        last_advantage = 0

        for t in reversed(range(len(rewards))):
            if t == len(rewards) - 1:
                next_value = 0
            else:
                next_value = values[t + 1]
            delta = rewards[t] + EXPERIMENT_CONFIG['GAMMA'] * next_value * (1 - dones[t]) - values[t]
            advantages[t] = delta + EXPERIMENT_CONFIG['GAMMA'] * EXPERIMENT_CONFIG['GAE_LAMBDA'] * (1 - dones[t]) * last_advantage
            last_advantage = advantages[t]

        returns = advantages + values
        advantages = (advantages - np.mean(advantages)) / (np.std(advantages) + 1e-8)
        return advantages, returns

    def get_batch_iterator(self, advantages, returns):
        indices = np.arange(len(self.states))
        np.random.shuffle(indices)
        batch_size = EXPERIMENT_CONFIG['PPO_BATCH_SIZE']
        for start in range(0, len(self.states), batch_size):
            end = min(start + batch_size, len(self.states))
            batch_indices = indices[start:end]
            yield {
                'states': torch.FloatTensor(np.array([self.states[i] for i in batch_indices])).to(device),
                'actions': torch.FloatTensor(np.array([self.actions[i] for i in batch_indices])).to(device),
                'old_log_probs': torch.FloatTensor(np.array([self.log_probs[i] for i in batch_indices])).to(device),
                'advantages': torch.FloatTensor(advantages[batch_indices]).to(device),
                'returns': torch.FloatTensor(returns[batch_indices]).to(device)
            }

    def get_top_k_per_path(self, num_paths, k=20):
        path_samples = {i: [] for i in range(num_paths)}
        for idx in range(len(self.states)):
            path_idx = self.path_indices[idx]
            similarity = self.similarities[idx]
            state = self.states[idx]
            path_samples[path_idx].append((idx, similarity, state))

        top_k_results = {}
        for path_idx in range(num_paths):
            samples = path_samples[path_idx]
            if len(samples) == 0:
                top_k_results[path_idx] = []
                continue
            samples.sort(key=lambda x: x[1], reverse=True)
            top_k = samples[:k]
            top_k_results[path_idx] = []
            for sample in top_k:
                normalized_state = sample[2]
                original_state = denormalize_state(normalized_state)
                original_state_int = np.round(original_state).astype(int)
                top_k_results[path_idx].append({
                    'state': original_state_int,
                    'similarity': sample[1],
                    'triggered': execute_Tr(*original_state_int)
                })
        return top_k_results

    def clear(self):
        self.states.clear()
        self.actions.clear()
        self.rewards.clear()
        self.values.clear()
        self.log_probs.clear()
        self.dones.clear()
        self.path_indices.clear()
        self.similarities.clear()

    def __len__(self):
        return len(self.states)

# === PPO Agent ===
class PPOAgent:
    def __init__(self, state_dim=3, action_dim=3):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.actor = PPOActor(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.actor_optimizer = optim.Adam(self.actor.parameters(), lr=EXPERIMENT_CONFIG['ACTOR_LR'])
        self.critic = PPOCritic(state_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_optimizer = optim.Adam(self.critic.parameters(), lr=EXPERIMENT_CONFIG['CRITIC_LR'])
        self.buffer = PPOBuffer()
        self.update_count = 0

    def get_action(self, state):
        min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
        max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
        normalized_state = (state - (min_vals + max_vals) / 2) / ((max_vals - min_vals) / 2)
        state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)
        with torch.no_grad():
            action, log_prob = self.actor.get_action_and_log_prob(state_tensor)
            value = self.critic(state_tensor)
        action = action.cpu().numpy()[0]
        log_prob = log_prob.cpu().item()
        value = value.cpu().item()
        return action, log_prob, value

    def store_experience(self, state, action, reward, value, log_prob, done, path_idx, similarity):
        min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
        max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
        normalized_state = (state - (min_vals + max_vals) / 2) / ((max_vals - min_vals) / 2)
        self.buffer.store(normalized_state, action, reward, value, log_prob, done, path_idx, similarity)

    def update(self):
        if len(self.buffer) == 0:
            return
        advantages, returns = self.buffer.compute_advantages()
        for epoch in range(EXPERIMENT_CONFIG['PPO_EPOCHS']):
            for batch in self.buffer.get_batch_iterator(advantages, returns):
                new_log_probs = self.actor.get_log_prob(batch['states'], batch['actions'])
                ratio = torch.exp(new_log_probs - batch['old_log_probs'])
                surr1 = ratio * batch['advantages']
                surr2 = torch.clamp(ratio, 1 - EXPERIMENT_CONFIG['CLIP_EPSILON'], 1 + EXPERIMENT_CONFIG['CLIP_EPSILON']) * batch['advantages']
                actor_loss = -torch.min(surr1, surr2).mean()
                self.actor_optimizer.zero_grad()
                actor_loss.backward()
                torch.nn.utils.clip_grad_norm_(self.actor.parameters(), 0.5)
                self.actor_optimizer.step()

                new_values = self.critic(batch['states'])
                critic_loss = F.mse_loss(new_values, batch['returns'])
                self.critic_optimizer.zero_grad()
                critic_loss.backward()
                torch.nn.utils.clip_grad_norm_(self.critic.parameters(), 0.5)
                self.critic_optimizer.step()

        self.update_count += 1
        self.buffer.clear()
        if self.update_count % 2 == 0:
            print(f"  -> PPO completed (Run {self.update_count})")

# === Metric 计算函数 ===
def calculate_run_performance(run_idx, ppo_results, training_time, total_steps, update_count, agent):
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)
    total_reward = 0
    all_similarities = []
    total_samples = 0
    all_rewards = []

    for path_idx in range(num_paths):
        samples = ppo_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']
            total_reward += reward
            all_rewards.append(reward)
            all_similarities.append(similarity)
            total_samples += 1

    average_reward = total_reward / total_samples if total_samples > 0 else 0
    convergence = np.mean(all_similarities) if all_similarities else 0
    environment_adaptability = 1 / (np.std(all_similarities) + 1e-8) if len(all_similarities) > 1 else 0
    generalization_ability = convergence
    computational_efficiency = total_steps / training_time if training_time > 0 else 0
    policy_update_frequency = update_count / training_time if training_time > 0 else 0
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0

    return {
        'Run': run_idx + 1,
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }

# === Excel 导出函数 ===
def export_to_excel(all_ppo_results, all_performance_data, target_paths, output_path="PPO_20_run.xlsx"):
    print("\n导出Excel...")
    all_ppo_summary_data = []
    all_ppo_detailed_data = []

    for run_idx, (ppo_results, performance_data) in enumerate(zip(all_ppo_results, all_performance_data)):
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = ppo_results[path_idx]
            if len(samples) == 0:
                all_ppo_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue
            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = 'Yes' if perfect_count > 0 else 'No'
            all_ppo_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })
            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']
                all_ppo_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Sample ID': sample_idx + 1,
                    'X(流量)': int(state[0]),
                    'Y(电压)': int(state[1]),
                    'Z(电流)': int(state[2]),
                    'Similarity': round(similarity, 4),
                    'Perfect Coverage': 'Yes' if abs(similarity - 1.0) < 0.001 else 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path))),
                    'Triggered Rules': ', '.join(map(str, sorted(triggered))),
                    'Hit Rule Count': len(target_path.intersection(triggered)),
                    'Target Rule Count': len(target_path)
                })

    dqn_summary_df = pd.DataFrame(all_ppo_summary_data)
    dqn_detailed_df = pd.DataFrame(all_ppo_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        dqn_summary_df.to_excel(writer, sheet_name='PPOPath', index=False)
        dqn_detailed_df.to_excel(writer, sheet_name='PPODetailed Sample Data', index=False)
        selected_columns = [
            'Run', 'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df[selected_columns].to_excel(writer, sheet_name='Metric', index=False)

        workbook = writer.book
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        for sheet_name in ['PPOPath', 'PPODetailed Sample Data', 'Metric']:
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

    print(f"Excel已保存: {output_path}")

# === 训练流程 ===
def train_ppo_workflow():
    print("=" * 80)
    print("PPO 训练 - 热电流量混合规则集")
    print("Similarity: 覆盖规则数 / 目标路径规则数")
    print("=" * 80)

    agent = PPOAgent()
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n生成路径样本: {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']} / 路径")
    path_samples = {}
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']

    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            state = np.array([
                np.random.randint(min_vals[0], max_vals[0] + 1),
                np.random.randint(min_vals[1], max_vals[1] + 1),
                np.random.randint(min_vals[2], max_vals[2] + 1)
            ], dtype=np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  路径 {path_idx + 1}/{num_paths}: {len(samples)} 个初始样本")

    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size

    print(f"\n批次大小: {batch_size}, 单样本迭代步数: {EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(f"总批次数: {num_batches} /路径 × {num_paths} 路径 = {num_batches * num_paths}")
    print("-" * 80)

    global_buffer = PPOBuffer()

    for batch_idx in range(num_batches):
        print(f"\n批次 {batch_idx + 1}/{num_batches}")

        for path_idx in range(num_paths):
            target_path = target_paths[path_idx]
            batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]

            batch_rewards = []
            batch_similarities = []

            for initial_state in batch_samples:
                state = initial_state.copy()
                episode_reward = 0
                final_similarity = 0

                for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                    action, log_prob, value = agent.get_action(state)
                    next_state = state + action
                    next_state = clip_state(next_state)

                    triggered = execute_Tr(*next_state)
                    reward = unified_reward_function(triggered, target_path)
                    similarity = coverage_similarity(triggered, target_path)
                    done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)

                    agent.store_experience(state, action, reward, value, log_prob, done, path_idx, similarity)
                    normalized_state = (state - (min_vals + max_vals) / 2) / ((max_vals - min_vals) / 2)
                    global_buffer.store(
                        normalized_state,
                        action, reward, value, log_prob, done, path_idx, similarity
                    )

                    state = next_state
                    episode_reward += reward
                    final_similarity = similarity
                    total_steps += 1

                batch_rewards.append(episode_reward)
                batch_similarities.append(final_similarity)

            avg_reward = np.mean(batch_rewards)
            avg_similarity = np.mean(batch_similarities)
            print(f"  路径 {path_idx + 1}: 平均奖励={avg_reward:.2f}, 平均相似度={avg_similarity:.4f}")

        print(f"\n  执行 PPO 策略更新...")
        agent.update()
        print(f"  全局缓冲区累计样本数: {len(global_buffer)}")

    training_time = time.time() - start_time

    print("\n" + "=" * 80)
    print(f"单次运行训练完成! 总耗时: {training_time:.2f} 秒, 总交互步数: {total_steps}")
    print(f"全局缓冲区大小: {len(global_buffer)}")
    print(f"PPO 策略更新次数: {agent.update_count}")
    print("=" * 80)

    print(f"\n提取每条路径的 Top-{EXPERIMENT_CONFIG['TOP_K_SAMPLES']} 高相似度样本...")
    ppo_top_k_results = global_buffer.get_top_k_per_path(num_paths, EXPERIMENT_CONFIG['TOP_K_SAMPLES'])

    return agent, ppo_top_k_results, training_time, total_steps, agent.update_count

# === 主程序 ===
def main():
    print("\n" + "=" * 80)
    print("PPO 多轮实验 - 热电流量混合规则测试")
    print("参数范围: X流量 20~80, Y电压 200~250, Z电流 4~20")
    print("=" * 80)

    all_ppo_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'='*80}")
        print(f"开始第 {run_idx + 1}/{EXPERIMENT_CONFIG['NUM_RUNS']} 次独立运行")
        print(f"{'='*80}")

        ppo_agent, ppo_results, training_time, total_steps, update_count = train_ppo_workflow()
        performance_data = calculate_run_performance(
            run_idx, ppo_results, training_time, total_steps, update_count, ppo_agent
        )

        all_ppo_results.append(ppo_results)
        all_performance_data.append(performance_data)

        print(f"\n第 {run_idx + 1} 次运行完成!")
        print(f"  总奖励: {performance_data['Total Reward']}")
        print(f"  平均奖励: {performance_data['Average Reward']}")
        print(f"  平均相似度: {performance_data['Average Similarity']}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"PPO_thermal_flow_20run_{timestamp}.xlsx"
    export_to_excel(all_ppo_results, all_performance_data, target_paths, output_path)

    print("\n" + "=" * 80)
    print(f"{EXPERIMENT_CONFIG['NUM_RUNS']} 次运行整体统计摘要")
    print("=" * 80)

    total_rewards = [p['Total Reward'] for p in all_performance_data]
    average_rewards = [p['Average Reward'] for p in all_performance_data]
    convergences = [p['Convergence'] for p in all_performance_data]
    environment_adaptabilities = [p['Environment Adaptability'] for p in all_performance_data]
    generalization_abilities = [p['Generalization Ability'] for p in all_performance_data]
    computational_efficiencies = [p['Computational Efficiency'] for p in all_performance_data]
    policy_update_frequencies = [p['Policy Update Frequency'] for p in all_performance_data]
    avg_similarities = [p['Average Similarity'] for p in all_performance_data]

    print(f"\n总奖励统计:")
    print(f"  均值: {np.mean(total_rewards):.2f}")
    print(f"  标准差: {np.std(total_rewards):.2f}")

    print(f"\n平均奖励统计:")
    print(f"  均值: {np.mean(average_rewards):.4f}")
    print(f"  标准差: {np.std(average_rewards):.4f}")

    print(f"\n收敛度统计 (平均相似度):")
    print(f"  均值: {np.mean(convergences):.4f}")
    print(f"  标准差: {np.std(convergences):.4f}")

    print(f"\n环境适应性统计:")
    print(f"  均值: {np.mean(environment_adaptabilities):.4f}")
    print(f"  标准差: {np.std(environment_adaptabilities):.4f}")

    print(f"\n泛化能力统计:")
    print(f"  均值: {np.mean(generalization_abilities):.4f}")
    print(f"  标准差: {np.std(generalization_abilities):.4f}")

    print(f"\n计算效率统计 (步/秒):")
    print(f"  均值: {np.mean(computational_efficiencies):.2f}")
    print(f"  标准差: {np.std(computational_efficiencies):.2f}")

    print(f"\n策略更新频率统计 (次/秒):")
    print(f"  均值: {np.mean(policy_update_frequencies):.4f}")
    print(f"  标准差: {np.std(policy_update_frequencies):.4f}")

    print(f"\n平均相似度统计:")
    print(f"  均值: {np.mean(avg_similarities):.4f}")
    print(f"  标准差: {np.std(avg_similarities):.4f}")

    print("\n" + "=" * 80)
    print(f" {EXPERIMENT_CONFIG['NUM_RUNS']} 次运行全部完成! 结果已导出至 Excel")
    print("=" * 80)

if __name__ == "__main__":
    main()