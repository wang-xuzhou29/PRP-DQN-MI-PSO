import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
import numpy as np
import random
import time
import math
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# === 配置（新范围，匹配 section3） ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'MIN_X': 20, 'MAX_X': 80,        # 流量范围 20~80
    'MIN_Y': 200, 'MAX_Y': 250,      # 电压范围 200~250
    'MIN_Z': 4, 'MAX_Z': 20,         # 电流范围 4~20
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 5,
    'NUM_ROUNDS': 5,
    'REPLAY_BATCH_SIZE': 64,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'LEARNING_RATE': 3e-4,
    'NUM_RUNS': 20,
    'TOP_K_SAMPLES': 20,
    'REPLAY_BUFFER_CAPACITY': 20000,
}

# === 目标路径组（与 section3 规则编号匹配） ===
targetPaths = [
    {4, 8, 13, 15, 18, 19, 21, 22, 24, 29, 30, 34, 37, 38, 42, 43, 44, 45, 46, 51, 52, 57, 68, 71, 72, 93, 94, 96, 97,
     98, 101, 102},
    {1, 2, 8, 13, 15, 18, 19, 21, 22, 24, 29, 30, 34, 37, 38, 42, 43, 44, 45, 46, 51, 52, 57, 68, 71, 72, 93, 94, 96,
     98, 101, 102},
    {4, 8, 10, 11, 13, 14, 15, 17, 19, 21, 22, 24, 29, 30, 34, 38, 42, 43, 44, 45, 46, 53, 55, 56, 58, 59, 71, 72, 73,
     76, 94, 96},
    {4, 7, 10, 11, 13, 14, 15, 17, 21, 22, 24, 29, 34, 37, 38, 42, 43, 44, 45, 46, 51, 52, 57, 71, 72, 73, 76, 77, 93,
     94, 96, 99},
    {4, 8, 13, 14, 15, 16, 19, 21, 22, 24, 29, 30, 34, 38, 42, 43, 44, 45, 46, 53, 55, 56, 59, 71, 72, 73, 76, 94, 96,
     103, 104},
    {4, 7, 13, 15, 16, 24, 29, 31, 32, 35, 36, 37, 38, 42, 43, 44, 45, 46, 51, 52, 57, 71, 72, 73, 76, 93, 94, 96, 99,
     103, 104},
    {7, 10, 11, 14, 15, 17, 24, 29, 31, 32, 35, 37, 38, 42, 43, 44, 45, 46, 53, 55, 56, 59, 71, 72, 73, 76, 77, 88, 89,
     94, 96},
    {3, 8, 9, 10, 11, 14, 15, 17, 21, 24, 29, 30, 34, 38, 42, 43, 44, 45, 46, 53, 55, 56, 58, 59, 71, 72, 73, 76, 77,
     94, 96},
    {3, 8, 9, 14, 15, 16, 19, 21, 24, 29, 30, 34, 38, 42, 43, 44, 45, 46, 53, 55, 56, 71, 72, 73, 76, 94, 96, 100, 103,
     104},
    {7, 10, 11, 13, 15, 17, 24, 28, 29, 31, 32, 35, 37, 38, 42, 43, 44, 45, 46, 51, 52, 57, 71, 72, 73, 76, 77, 93, 94,
     96},
    {7, 15, 16, 24, 28, 29, 31, 32, 35, 36, 37, 38, 42, 43, 44, 45, 46, 52, 57, 71, 72, 73, 76, 92, 95, 100, 103, 104},
    {1, 2, 13, 18, 19, 21, 22, 24, 29, 33, 38, 42, 43, 44, 45, 46, 53, 55, 56, 59, 68, 71, 72, 94, 96, 98, 101, 102},
    {1, 2, 8, 15, 18, 20, 23, 25, 30, 34, 36, 37, 38, 42, 43, 44, 45, 46, 51, 52, 57, 68, 71, 72, 93, 94, 96, 101},
    {1, 2, 5, 6, 12, 13, 15, 18, 20, 23, 25, 30, 34, 36, 37, 38, 42, 43, 47, 51, 52, 57, 69, 70, 93, 94, 96, 101},
    {4, 13, 24, 29, 33, 39, 40, 44, 45, 46, 53, 55, 56, 58, 59, 68, 71, 72, 86, 87, 90, 94, 96, 97, 98, 101, 102},
    {1, 2, 24, 29, 33, 39, 40, 44, 45, 46, 53, 55, 56, 58, 59, 68, 71, 72, 85, 86, 87, 90, 94, 96, 101, 102},
    {8, 10, 11, 14, 15, 17, 26, 27, 34, 42, 43, 44, 45, 46, 53, 55, 56, 58, 59, 71, 72, 73, 76, 77, 94, 96},
    {11, 13, 17, 24, 28, 29, 32, 35, 36, 37, 38, 42, 43, 44, 45, 46, 51, 52, 57, 74, 75, 93, 94, 96, 99},
    {11, 15, 18, 24, 28, 29, 31, 32, 35, 42, 43, 48, 49, 50, 53, 55, 56, 58, 59, 74, 75, 88, 89, 94, 96},
    {4, 13, 16, 24, 29, 33, 40, 44, 45, 46, 54, 57, 71, 72, 73, 86, 87, 90, 94, 96, 97, 98, 103, 104},
    {1, 2, 13, 24, 29, 35, 41, 44, 45, 46, 60, 61, 62, 63, 65, 66, 67, 68, 71, 72, 91, 92, 95, 101},
    {4, 7, 13, 16, 24, 29, 35, 41, 44, 45, 46, 57, 64, 71, 72, 73, 91, 92, 95, 97, 98, 103, 104},
    {1, 2, 12, 25, 33, 36, 37, 38, 42, 43, 47, 51, 52, 57, 70, 82, 84, 85, 86, 87, 90, 94, 96},
    {1, 2, 12, 23, 25, 33, 38, 42, 43, 47, 53, 55, 56, 59, 70, 81, 82, 84, 86, 87, 90, 96},
    {12, 15, 25, 33, 41, 47, 60, 61, 62, 63, 65, 66, 67, 70, 79, 80, 83, 84, 92, 95},
    {12, 23, 25, 32, 35, 41, 47, 57, 60, 61, 62, 63, 65, 66, 67, 69, 70, 78, 92, 95},
]

# === 工具函数（使用配置中的范围） ===
def get_bounds():
    mins = np.array([EXPERIMENT_CONFIG['MIN_X'], EXPERIMENT_CONFIG['MIN_Y'], EXPERIMENT_CONFIG['MIN_Z']])
    maxs = np.array([EXPERIMENT_CONFIG['MAX_X'], EXPERIMENT_CONFIG['MAX_Y'], EXPERIMENT_CONFIG['MAX_Z']])
    return mins, maxs

def clip_state(state):
    mins, maxs = get_bounds()
    return np.clip(state, mins, maxs)

def denormalize_state(normalized_state):
    mins, maxs = get_bounds()
    return normalized_state * (maxs - mins) / 2 + (mins + maxs) / 2

def coverage_similarity(triggered, target_path):
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0
    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)

def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']
    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']
    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']
    return reward

# ==================== section3 触发函数（顶层定义） ====================
def section3_thermal_electrical_flow_hybrid(x, y, z):
    triggered = set()

    # 第1组: 流量和电压的复合条件 (分支1-4)
    if (x > 60 and x < 70 and y > 225 and y < 235) != (x > 40 and x < 70 and y > 225 and y < 235):
        triggered.add(1)
    if (x > 60 and x < 70 and y > 225 and y < 235) != (x > 20 and x < 70 and y > 225 and y < 235):
        triggered.add(2)
    if (x > 60 and x < 70 and y > 225 and y < 235) != (x > 60 and x < 70 and y > 220 and y < 235):
        triggered.add(3)
    if (x > 60 and x < 70 and y > 225 and y < 235) != (x > 60 and x < 70 and y > 225 and y < 230):
        triggered.add(4)

    # 第2组: 流量和电流的复合条件 (分支5-8)
    if (x > 55 and x < 75 and z > 11 and z < 14) != (x > 45 and x < 75 and z > 11 and z < 14):
        triggered.add(5)
    if (x > 55 and x < 75 and z > 11 and z < 14) != (x > 35 and x < 75 and z > 11 and z < 14):
        triggered.add(6)
    if (x > 55 and x < 75 and z > 11 and z < 14) != (x > 55 and x < 75 and z > 4 and z < 14):
        triggered.add(7)
    if (x > 55 and x < 75 and z > 11 and z < 14) != (x > 55 and x < 75 and z > 11 and z < 4):
        triggered.add(8)

    # 第3组: 电压和电流的复合条件 (分支9)
    if (y > 225 and y < 235 and z > 11 and z < 14) != (y > 220 and y < 235 and z > 11 and z < 14):
        triggered.add(9)

    # 第4组: 流量精确区间判断 (分支10-12)
    if (x > 55 and x < 67) != (x > 55 and x < 75):
        triggered.add(10)
    if (x > 55 and x < 67) != (x > 55 and x < 80):
        triggered.add(11)
    if (x > 55 and x < 67) != (x > 35 and x < 67):
        triggered.add(12)

    # 第5组: 电压和电流的精确条件 (分支13-18)
    if (y > 228 and y < 232) != (y > 228 and y < 230):
        triggered.add(13)
    if (abs(z - 12.5) < 0.8) != (abs(z - 11.5) < 0.8):
        triggered.add(14)
    if (abs(z - 12.5) < 0.8) != (abs(z - 10.5) < 0.8):
        triggered.add(15)
    if ((x - 50) / 30 > 0.45 and (x - 50) / 30 < 0.55) != ((x - 50) / 30 > 0.45 and (x - 50) / 20 < 0.55):
        triggered.add(16)
    if ((x - 50) / 30 > 0.45 and (x - 50) / 30 < 0.55) != ((x - 50) / 30 > 0.45 and (x - 50) / 30 < 0.85):
        triggered.add(17)
    if (abs(z - 12.5) < 0.8) != (abs(z - 13.5) < 0.8):
        triggered.add(18)

    # 第6组: 三参数复合条件 (分支19-22)
    if (x > 58 and x < 72 and y > 223 and y < 237 and z > 10.5 and z < 14.5) != (x > 68 and x < 72 and y > 223 and y < 237 and z > 10.5 and z < 14.5):
        triggered.add(19)
    if (x > 58 and x < 72 and y > 223 and y < 237 and z > 10.5 and z < 14.5) != (x > 48 and x < 72 and y > 223 and y < 237 and z > 10.5 and z < 14.5):
        triggered.add(20)
    if (x > 58 and x < 72 and y > 223 and y < 237 and z > 10.5 and z < 14.5) != (x > 58 and x < 72 and y > 233 and y < 237 and z > 10.5 and z < 14.5):
        triggered.add(21)
    if (x > 58 and x < 72 and y > 223 and y < 237 and z > 10.5 and z < 14.5) != (x > 58 and x < 72 and y > 223 and y < 227 and z > 10.5 and z < 14.5):
        triggered.add(22)

    # 第7组: 比值计算条件 (分支23-29)
    if y != 0:
        if (x / (y / 10) > 2.5 and x / (y / 10) < 3.5) != (x / (y / 13) > 2.5 and x / (y / 10) < 3.5):
            triggered.add(23)
        if (x / (y / 10) > 2.5 and x / (y / 10) < 3.5) != (x / (y / 6) > 2.5 and x / (y / 10) < 3.5):
            triggered.add(24)
        if (x / (y / 10) > 2.5 and x / (y / 10) < 3.5) != (x / (y / 10) > 1.5 and x / (y / 10) < 3.5):
            triggered.add(25)
        if (x / (y / 10) > 2.5 and x / (y / 10) < 3.5) != (x / (y / 10) > 2.5 and x / (y / 10) < 5.5):
            triggered.add(26)
        if (x / (y / 10) > 2.5 and x / (y / 10) < 3.5) != (x / (y / 10) > 2.5 and x / ((y + 100) / 10) < 3.5):
            triggered.add(27)
        if (x / (y / 10) > 2.5 and x / (y / 10) < 3.5) != (x / (y / 10) > 2.5 and x / (y / 11) < 3.5):
            triggered.add(28)
        if (x / (y / 10) > 2.5 and x / (y / 10) < 3.5) != (x / (y / 10) > 2.5 and x / (y / 15) < 3.5):
            triggered.add(29)

    # 第8组: 流量电流比值条件 (分支30-35)
    if z != 0:
        if (x / z > 4.5 and x / z < 6.5) != (x / z > 5.5 and x / z < 6.5):
            triggered.add(30)
        if (x / z > 4.5 and x / z < 6.5) != (x / z > 4.5 and x / z < 7.5):
            triggered.add(31)
        if (x / z > 4.5 and x / z < 6.5) != (x / z > 4.5 and x / z < 8.5):
            triggered.add(32)
        if (x / z > 4.5 and x / z < 6.5) != (y / z > 4.5 and x / z < 6.5):
            triggered.add(33)
        if y != 0:
            if (x / z > 4.5 and x / z < 6.5) != (x / (y / 10) > 4.5 and x / z < 6.5):
                triggered.add(34)
            if (x / z > 4.5 and x / z < 6.5) != (x / z > 4.5 and x / (y / 10) < 6.5):
                triggered.add(35)

    # 第9组: 乘积条件 (分支36-43)
    if (x * z > 600 and x * z < 1000) != (x * z > 700 and x * z < 1000):
        triggered.add(36)
    if (x * z > 600 and x * z < 1000) != (x * z > 800 and x * z < 1000):
        triggered.add(37)
    if (x * z > 600 and x * z < 1000) != (x * z > 900 and x * z < 1000):
        triggered.add(38)
    if (x * z > 600 and x * z < 1000) != (x * z > 600 and x * z < 1100):
        triggered.add(39)
    if (x * z > 600 and x * z < 1000) != (x * z > 600 and x * z < 1200):
        triggered.add(40)
    if (x * z > 600 and x * z < 1000) != (x * y > 600 and x * z < 1000):
        triggered.add(41)
    if (x * z > 600 and x * z < 1000) != (x * z > 600 and x * y < 1000):
        triggered.add(42)
    if (x * z > 600 and x * z < 1000) != (x * z > 600 and y * z < 1000):
        triggered.add(43)

    # 第10组: 加权平均条件 (分支44-50)
    if ((x + y / 10) / 2 > 40 and (x + y / 10) / 2 < 50) != ((x + z / 10) / 2 > 40 and (x + y / 10) / 2 < 50):
        triggered.add(44)
    if ((x + y / 10) / 2 > 40 and (x + y / 10) / 2 < 50) != ((z + y / 10) / 2 > 40 and (x + y / 10) / 2 < 50):
        triggered.add(45)
    if ((x + y / 10) / 2 > 40 and (x + y / 10) / 2 < 50) != ((x + y / 10) / 3 > 40 and (x + y / 10) / 2 < 50):
        triggered.add(46)
    if ((x + y / 10) / 2 > 40 and (x + y / 10) / 2 < 50) != ((x + y / 10) / 2 > 30 and (x + y / 10) / 2 < 50):
        triggered.add(47)
    if ((x + y / 10) / 2 > 40 and (x + y / 10) / 2 < 50) != ((x + y / 10) / 2 > 40 and (x + y / 14) / 2 < 50):
        triggered.add(48)
    if ((x + y / 10) / 2 > 40 and (x + y / 10) / 2 < 50) != ((x + y / 10) / 2 > 40 and (x + y / 10) / 3 < 50):
        triggered.add(49)
    if ((x + y / 10) / 2 > 40 and (x + y / 10) / 2 < 50) != ((x + y / 10) / 2 > 40 and (x + y / 10) / 2 < 60):
        triggered.add(50)

    # 第11组: 复杂乘积条件 (分支51-59)
    if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != ((x * (y / 10) * z) > 15000 and (x * (y / 10) * z) < 26000):
        triggered.add(51)
    if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != ((x * (y / 5) * z) > 18000 and (x * (y / 10) * z) < 26000):
        triggered.add(52)
    if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != ((x * (y / 10) * z) > 18000 and (x * (y / 5) * z) < 26000):
        triggered.add(53)
    if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != ((x * (y / 10) * z) > 18000 and (z * (y / 10) * z) < 26000):
        triggered.add(54)
    if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != ((z * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000):
        triggered.add(55)
    if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != ((x * z * z) > 18000 and (x * (y / 10) * z) < 26000):
        triggered.add(56)
    if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != ((x * x * z) > 18000 and (x * z * z) < 26000):
        triggered.add(57)
    if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 20000):
        triggered.add(58)
    if ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 26000) != ((x * (y / 10) * z) > 18000 and (x * (y / 10) * z) < 19000):
        triggered.add(59)

    # 第12组: 幂次条件 (分支60-67)
    if (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4500) != (x ** 2 + (y / 10) ** 3 + (z * 5) ** 2 > 4500):
        triggered.add(60)
    if (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4500) != (x ** 3 + (y / 10) ** 2 + (z * 5) ** 2 > 4500):
        triggered.add(61)
    if (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4500) != (x ** 2 + (y / 10) ** 2 + (z * 5) ** 3 > 4500):
        triggered.add(62)
    if (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4500) != (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 3500):
        triggered.add(63)
    if (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4500) != (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 5500):
        triggered.add(64)
    if (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4500) != (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4000):
        triggered.add(65)
    if (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4500) != (x ** 2 + (y / 10) ** 2 + (z * 10) ** 2 > 4500):
        triggered.add(66)
    if (x ** 2 + (y / 10) ** 2 + (z * 5) ** 2 > 4500) != (x ** 2 + (y / 5) ** 2 + (z * 5) ** 2 > 4500):
        triggered.add(67)

    # 第13组: 归一化条件 (分支68-77)
    if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 65 > 0.95 and x / 65 < 1.15):
        triggered.add(68)
    if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 55 > 0.85 and x / 65 < 1.15):
        triggered.add(69)
    if (x / 65 > 0.85 and x / 65 < 1.15) != (y / 65 > 0.85 and x / 65 < 1.15):
        triggered.add(70)
    if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 65 > 0.85 and y / 65 < 1.15):
        triggered.add(71)
    if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 65 > 0.85 and x / 45 < 1.15):
        triggered.add(72)
    if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 65 > 0.85 and x / 55 < 1.15):
        triggered.add(73)
    if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 65 > 0.85 and x / 75 < 1.15):
        triggered.add(74)
    if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 65 > 0.85 and x / 65 < 2.15):
        triggered.add(75)
    if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 65 > 0.85 and x / 65 < 1.0):
        triggered.add(76)
    if (x / 65 > 0.85 and x / 65 < 1.15) != (x / 65 > 0.85 and x / 65 < 1.05):
        triggered.add(77)

    # 第14组: 阈值条件 (分支78-84)
    if (x < 45 and y < 220) != (x < 55 and y < 220):
        triggered.add(78)
    if (x < 45 and y < 220) != (x < 35 and y < 220):
        triggered.add(79)
    if (x < 45 and y < 220) != (x < 25 and y < 220):
        triggered.add(80)
    if (x < 45 and y < 220) != (x < 45 and y < 230):
        triggered.add(81)
    if (x < 45 and y < 220) != (x < 45 and y < 240):
        triggered.add(82)
    if (x < 45 and y < 220) != (x < 45 and y < 210):
        triggered.add(83)
    if (x < 45 and y < 220) != (x < 45 and y > 220):
        triggered.add(84)

    # 第15组: 高阈值条件 (分支85-90)
    if (y > 240 and z > 16) != (y > 230 and z > 16):
        triggered.add(85)
    if (y > 240 and z > 16) != (y > 220 and z > 16):
        triggered.add(86)
    if (y > 240 and z > 16) != (y > 210 and z > 16):
        triggered.add(87)
    if (y > 240 and z > 16) != (y > 240 and x > 16):
        triggered.add(88)
    if (y > 240 and z > 16) != (y > 240 and z < 16):
        triggered.add(89)
    if (y > 240 and z > 16) != (y < 240 and z > 16):
        triggered.add(90)

    # 第16组: 复合逻辑条件 (分支91-96)
    if ((x + y / 10 + z * 5) < 95 or (x + y / 10 + z * 5) > 135) != ((x + z + z * 5) < 95 or (x + y / 10 + z * 5) > 135):
        triggered.add(91)
    if ((x + y / 10 + z * 5) < 95 or (x + y / 10 + z * 5) > 135) != ((z + y / 10 + z * 5) < 95 or (x + y / 10 + z * 5) > 135):
        triggered.add(92)
    if ((x + y / 10 + z * 5) < 95 or (x + y / 10 + z * 5) > 135) != ((x + y / 10 + z * 5) < 95 or (x + z + z * 5) > 135):
        triggered.add(93)
    if ((x + y / 10 + z * 5) < 95 or (x + y / 10 + z * 5) > 135) != ((x + y / 10 + z * 5) < 95 or (z + y / 10 + z * 5) > 135):
        triggered.add(94)
    if ((x + y / 10 + z * 5) < 95 or (x + y / 10 + z * 5) > 135) != ((x + y / 10 + z * 5) < 95 or (x + y / 10 + x * 5) > 135):
        triggered.add(95)
    if ((x + y / 10 + z * 5) < 95 or (x + y / 10 + z * 5) > 135) != ((x + y / 10 + z * 5) < 95 and (x + y / 10 + z * 5) > 135):
        triggered.add(96)

    # 第17组: 精确区间条件 (分支97-100)
    if (x > 64 and x < 66 and y > 229 and y < 231) != (x > 60 and x < 66 and y > 229 and y < 231):
        triggered.add(97)
    if (x > 64 and x < 66 and y > 229 and y < 231) != (x > 54 and x < 66 and y > 229 and y < 231):
        triggered.add(98)
    if (x > 64 and x < 66 and y > 229 and y < 231) != (x > 64 and x < 76 and y > 229 and y < 231):
        triggered.add(99)
    if (x > 64 and x < 66 and y > 229 and y < 231) != (x > 64 and x < 66 and y < 229 and y < 231):
        triggered.add(100)

    # 第18组: 标准化条件 (分支101-104)
    if ((x - 50) / 30 > 0.45 and (x - 50) / 30 < 0.55) != ((x - 40) / 30 > 0.45 and (x - 50) / 30 < 0.55):
        triggered.add(101)
    if ((x - 50) / 30 > 0.45 and (x - 50) / 30 < 0.55) != ((x - 50) / 20 > 0.45 and (x - 50) / 30 < 0.55):
        triggered.add(102)
    if ((x - 50) / 30 > 0.45 and (x - 50) / 30 < 0.55) != ((x - 50) / 30 > 0.65 and (x - 50) / 30 < 0.55):
        triggered.add(103)
    if ((x - 50) / 30 > 0.45 and (x - 50) / 30 < 0.55) != ((x - 50) / 30 > 0.45 and (x - 40) / 30 < 0.55):
        triggered.add(104)

    return triggered

# ========== 设置执行函数为 section3 ==========
execute_Tr = section3_thermal_electrical_flow_hybrid

# === DQN 网络 ===
class DQNNetwork(nn.Module):
    def __init__(self, action_size=30):
        super(DQNNetwork, self).__init__()
        hidden_dim = EXPERIMENT_CONFIG['HIDDEN_DIM']
        self.fc1 = nn.Linear(3, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)
        self.output = nn.Linear(hidden_dim, action_size)

    def forward(self, x):
        x = torch.relu(self.fc1(x))
        x = torch.relu(self.fc2(x))
        x = torch.relu(self.fc3(x))
        return self.output(x)

# === PathReplayBuffer ===
class PathReplayBuffer:
    def __init__(self, path_idx, capacity=20000):
        self.path_idx = path_idx
        self.capacity = capacity
        self.buffer = deque(maxlen=capacity)
        self.similarities = deque(maxlen=capacity)

    def push(self, state, action, reward, next_state, done, similarity):
        self.buffer.append((state, action, reward, next_state, done))
        self.similarities.append(similarity)

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return None
        batch = random.sample(self.buffer, batch_size)
        state, action, reward, next_state, done = map(np.stack, zip(*batch))
        return (
            torch.FloatTensor(state).to(device),
            torch.LongTensor(action).to(device),
            torch.FloatTensor(reward).to(device),
            torch.FloatTensor(next_state).to(device),
            torch.BoolTensor(done).to(device)
        )

    def get_top_k(self, k=20):
        if len(self.buffer) == 0:
            return []
        samples_with_sim = list(zip(self.buffer, self.similarities))
        samples_with_sim.sort(key=lambda x: x[1], reverse=True)
        top_k = samples_with_sim[:k]
        results = []
        for (state, _, _, _, _), similarity in top_k:
            original_state = denormalize_state(state)
            original_state_int = np.round(original_state).astype(int)
            results.append({
                'state': original_state_int,
                'similarity': similarity,
                'triggered': execute_Tr(*original_state_int)
            })
        return results

    def __len__(self):
        return len(self.buffer)

# === ImprovedDQNAgent ===
class ImprovedDQNAgent:
    def __init__(self, num_paths, action_size=30):
        self.action_size = action_size
        self.num_paths = num_paths
        self.epsilon = 0.9
        self.epsilon_min = 0.1
        self.epsilon_decay = 0.995

        self.q_network = DQNNetwork(action_size).to(device)
        self.target_network = DQNNetwork(action_size).to(device)
        self.optimizer = optim.Adam(self.q_network.parameters(), lr=EXPERIMENT_CONFIG['LEARNING_RATE'])

        self.replay_buffers = {}
        for path_idx in range(num_paths):
            self.replay_buffers[path_idx] = PathReplayBuffer(path_idx, EXPERIMENT_CONFIG['REPLAY_BUFFER_CAPACITY'])

        self.replay_train_count = 0
        self.update_target_network()

    def discrete_to_action_delta(self, action_idx):
        delta_values = [5, 3, 2, 1, 0.5, -0.5, -1, -2, -3, -5]
        if action_idx >= 30:
            action_idx = action_idx % 30
        dim = action_idx // 10
        delta_idx = action_idx % 10
        delta = delta_values[delta_idx]
        action_delta = np.zeros(3)
        action_delta[dim] = delta
        return action_delta

    def get_action(self, state):
        mins, maxs = get_bounds()
        normalized_state = (state - (mins + maxs) / 2) / ((maxs - mins) / 2)
        if random.random() < self.epsilon:
            action_idx = random.randint(0, self.action_size - 1)
        else:
            state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)
            with torch.no_grad():
                q_values = self.q_network(state_tensor)
                action_idx = q_values.argmax().item()
        action_delta = self.discrete_to_action_delta(action_idx)
        return action_delta, action_idx

    def store_experience(self, path_idx, state, action_idx, reward, next_state, done, similarity):
        mins, maxs = get_bounds()
        normalized_state = (state - (mins + maxs) / 2) / ((maxs - mins) / 2)
        normalized_next_state = (next_state - (mins + maxs) / 2) / ((maxs - mins) / 2)
        self.replay_buffers[path_idx].push(
            normalized_state, action_idx, reward,
            normalized_next_state, done, similarity
        )

    def replay_train(self, path_idx):
        batch = self.replay_buffers[path_idx].sample(EXPERIMENT_CONFIG['REPLAY_BATCH_SIZE'])
        if batch is None:
            return
        states, actions, rewards, next_states, dones = batch
        current_q_values = self.q_network(states).gather(1, actions.unsqueeze(1))
        with torch.no_grad():
            next_q_values = self.target_network(next_states).max(1)[0]
            target_q_values = rewards + (0.99 * next_q_values * ~dones)
        loss = F.mse_loss(current_q_values.squeeze(), target_q_values)
        self.optimizer.zero_grad()
        loss.backward()
        torch.nn.utils.clip_grad_norm_(self.q_network.parameters(), 1.0)
        self.optimizer.step()
        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay
        self.replay_train_count += 1
        if self.replay_train_count % 2 == 0:
            self.update_target_network()

    def update_target_network(self):
        self.target_network.load_state_dict(self.q_network.state_dict())

    def get_all_top_k(self, k=20):
        results = {}
        for path_idx in range(self.num_paths):
            results[path_idx] = self.replay_buffers[path_idx].get_top_k(k)
        return results

    def get_buffer_stats(self):
        stats = {}
        for path_idx in range(self.num_paths):
            stats[path_idx] = len(self.replay_buffers[path_idx])
        return stats

# === 性能计算 ===
def calculate_run_performance(run_idx, dqn_results, training_time, total_steps, update_count, agent):
    # 直接使用全局 targetPaths
    target_paths = targetPaths
    num_paths = len(target_paths)
    all_similarities = []
    total_reward = 0
    total_samples = 0
    all_rewards = []
    for path_idx in range(num_paths):
        samples = dqn_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']
            total_reward += reward
            all_rewards.append(reward)
            all_similarities.append(similarity)
            total_samples += 1
    average_reward = total_reward / total_samples if total_samples > 0 else 0
    convergence = np.mean(all_similarities) if all_similarities else 0
    environment_adaptability = 1 / (np.std(all_similarities) + 1e-8) if len(all_similarities) > 1 else 0
    generalization_ability = convergence
    computational_efficiency = total_steps / training_time if training_time > 0 else 0
    policy_update_frequency = update_count / training_time if training_time > 0 else 0
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0
    return {
        'Run': run_idx + 1,
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }

# === Excel导出 ===
def export_to_excel(all_dqn_results, all_performance_data, target_paths, output_path="DQN_20_run.xlsx"):
    print("\n导出Excel...")
    all_dqn_summary_data = []
    all_dqn_detailed_data = []
    for run_idx, (dqn_results, performance_data) in enumerate(zip(all_dqn_results, all_performance_data)):
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = dqn_results[path_idx]
            if len(samples) == 0:
                all_dqn_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue
            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = 'Yes' if perfect_count > 0 else 'No'
            all_dqn_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })
            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']
                all_dqn_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Sample ID': sample_idx + 1,
                    'X': int(state[0]),
                    'Y': int(state[1]),
                    'Z': int(state[2]),
                    'Similarity': round(similarity, 4),
                    'Perfect Coverage': 'Yes' if abs(similarity - 1.0) < 0.001 else 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path))),
                    'Triggered Rules': ', '.join(map(str, sorted(triggered))),
                    'Hit Rule Count': len(target_path.intersection(triggered)),
                    'Target Rule Count': len(target_path)
                })
    dqn_summary_df = pd.DataFrame(all_dqn_summary_data)
    dqn_detailed_df = pd.DataFrame(all_dqn_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        dqn_summary_df.to_excel(writer, sheet_name='DQNPath', index=False)
        dqn_detailed_df.to_excel(writer, sheet_name='DQNDetailed Sample Data', index=False)
        selected_columns = [
            'Run', 'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df[selected_columns].to_excel(writer, sheet_name='Metric', index=False)
        workbook = writer.book
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        for sheet_name in ['DQNPath', 'DQNDetailed Sample Data', 'Metric']:
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            # 可在此调整列宽
    print(f"Excel已保存: {output_path}")

# === 训练流程 ===
def train_dqn_workflow():
    print("=" * 80)
    print("开始DQN训练")
    print(f"路径数: {len(targetPaths)}, 每路径样本数: {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']}")
    print("=" * 80)
    num_paths = len(targetPaths)
    agent = ImprovedDQNAgent(num_paths=num_paths)
    start_time = time.time()
    total_steps = 0

    # 生成初始样本
    path_samples = {}
    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            state = np.array([
                random.randint(EXPERIMENT_CONFIG['MIN_X'], EXPERIMENT_CONFIG['MAX_X']),
                random.randint(EXPERIMENT_CONFIG['MIN_Y'], EXPERIMENT_CONFIG['MAX_Y']),
                random.randint(EXPERIMENT_CONFIG['MIN_Z'], EXPERIMENT_CONFIG['MAX_Z'])
            ], dtype=np.float32)
            samples.append(state)
        path_samples[path_idx] = samples

    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size
    num_rounds = EXPERIMENT_CONFIG['NUM_ROUNDS']

    for path_idx in range(num_paths):
        target_path = targetPaths[path_idx]
        print(f"\n开始训练路径 {path_idx+1}/{num_paths}, 目标规则数: {len(target_path)}")
        for round_idx in range(num_rounds):
            print(f"  轮次 {round_idx+1}/{num_rounds}")
            for batch_idx in range(num_batches):
                batch_samples = path_samples[path_idx][batch_idx*batch_size:(batch_idx+1)*batch_size]
                batch_rewards = []
                batch_similarities = []
                for initial_state in batch_samples:
                    state = initial_state.copy()
                    episode_reward = 0
                    final_similarity = 0
                    for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                        action_delta, action_idx = agent.get_action(state)
                        next_state = state + action_delta
                        next_state = clip_state(next_state)
                        triggered = execute_Tr(*next_state)
                        reward = unified_reward_function(triggered, target_path)
                        similarity = coverage_similarity(triggered, target_path)
                        done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)
                        agent.store_experience(path_idx, state, action_idx, reward, next_state, done, similarity)
                        state = next_state
                        episode_reward += reward
                        final_similarity = similarity
                        total_steps += 1
                    batch_rewards.append(episode_reward)
                    batch_similarities.append(final_similarity)
                avg_reward = np.mean(batch_rewards)
                avg_sim = np.mean(batch_similarities)
                print(f"    批次 {batch_idx+1}/{num_batches}: 平均奖励={avg_reward:.2f}, 平均相似度={avg_sim:.4f}, epsilon={agent.epsilon:.3f}")
                agent.replay_train(path_idx)
    training_time = time.time() - start_time
    print(f"\n训练完成，总耗时: {training_time:.2f}秒，总步数: {total_steps}")
    dqn_top_k_results = agent.get_all_top_k(EXPERIMENT_CONFIG['TOP_K_SAMPLES'])
    return agent, dqn_top_k_results, training_time, total_steps, agent.replay_train_count

# === 主程序 ===
def main():
    print("=" * 80)
    print("DQN - 20次运行实验")
    print(f"状态空间范围: X[20,80], Y[200,250], Z[4,20]")
    print("=" * 80)
    all_dqn_results = []
    all_performance_data = []
    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n开始第 {run_idx+1}/{EXPERIMENT_CONFIG['NUM_RUNS']} 次运行")
        dqn_agent, dqn_results, training_time, total_steps, update_count = train_dqn_workflow()
        performance_data = calculate_run_performance(run_idx, dqn_results, training_time, total_steps, update_count, dqn_agent)
        all_dqn_results.append(dqn_results)
        all_performance_data.append(performance_data)
        print(f"运行 {run_idx+1} 完成: 总奖励={performance_data['Total Reward']}, 收敛度={performance_data['Convergence']}")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"DQN_20_run_{timestamp}.xlsx"
    export_to_excel(all_dqn_results, all_performance_data, targetPaths, output_path)
    print("\n20次运行全部完成！")

if __name__ == "__main__":
    main()