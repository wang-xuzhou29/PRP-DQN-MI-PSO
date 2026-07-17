import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
import numpy as np
import random
import time
import math
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# === 配置（新范围，匹配 section6） ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'MIN_X': 1, 'MAX_X': 100,        # 温度范围 1~100
    'MIN_Y': 1, 'MAX_Y': 100,        # 电压范围 1~100
    'MIN_Z': 1, 'MAX_Z': 60,         # 流量范围 1~60
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 5,
    'NUM_ROUNDS': 5,
    'REPLAY_BATCH_SIZE': 64,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'LEARNING_RATE': 3e-4,
    'NUM_RUNS': 20,
    'TOP_K_SAMPLES': 20,
    'REPLAY_BUFFER_CAPACITY': 20000,
}

# === 目标路径组（与 section6 规则编号 1~109 匹配） ===
targetPaths = [
    {4, 6, 11, 12, 13, 16, 17, 18, 21, 24, 25, 26, 36, 39, 41, 43, 48, 51, 52, 54, 55, 58, 60, 64, 68, 69, 70, 71, 72,
     73, 74, 75, 81, 82, 83, 85, 86, 87, 89, 91, 93, 99, 102},
    {6, 10, 12, 13, 14, 15, 16, 17, 18, 20, 22, 25, 26, 36, 41, 45, 46, 47, 49, 51, 52, 54, 55, 57, 59, 60, 64, 69, 70,
     71, 72, 76, 81, 82, 86, 87, 88, 89, 91, 92, 98, 102},
    {6, 7, 11, 12, 13, 14, 15, 16, 17, 21, 24, 25, 26, 36, 39, 41, 43, 48, 49, 51, 52, 54, 55, 58, 60, 64, 67, 68, 72,
     75, 82, 86, 87, 89, 92, 97, 100, 102, 104},
    {6, 11, 12, 13, 14, 15, 16, 17, 19, 21, 24, 25, 26, 36, 48, 49, 53, 56, 57, 61, 64, 68, 69, 70, 71, 72, 73, 74, 76,
     78, 81, 82, 83, 86, 87, 89, 91, 92, 101},
    {6, 7, 13, 14, 15, 16, 17, 20, 22, 24, 25, 26, 30, 36, 39, 48, 49, 51, 52, 53, 56, 58, 59, 60, 64, 66, 68, 72, 75,
     81, 82, 86, 87, 88, 89, 92, 97, 100, 104},
    {4, 6, 12, 13, 16, 17, 19, 21, 22, 24, 25, 26, 33, 36, 39, 41, 43, 47, 51, 52, 54, 55, 60, 64, 69, 72, 73, 74, 76,
     78, 81, 82, 84, 85, 86, 89, 93, 94, 100},
    {5, 7, 11, 12, 13, 14, 15, 21, 22, 24, 25, 26, 33, 36, 39, 47, 51, 52, 53, 54, 55, 58, 59, 64, 69, 70, 71, 72, 75,
     82, 85, 86, 87, 88, 89, 91, 92, 97, 100},
    {6, 10, 14, 15, 16, 17, 18, 20, 22, 24, 25, 26, 31, 36, 49, 51, 52, 53, 55, 58, 59, 61, 64, 69, 73, 74, 75, 81, 82,
     83, 85, 86, 87, 89, 93, 94, 99, 100},
    {5, 7, 11, 13, 14, 15, 21, 23, 24, 25, 26, 30, 36, 39, 48, 51, 53, 54, 55, 58, 59, 60, 64, 66, 71, 72, 75, 82, 85,
     86, 87, 88, 89, 91, 92, 97, 99, 102},
    {6, 10, 12, 13, 15, 16, 17, 19, 23, 25, 26, 36, 41, 46, 50, 51, 52, 54, 55, 58, 59, 60, 64, 71, 72, 75, 81, 82, 86,
     87, 88, 89, 91, 92, 98, 99, 100},
    {4, 6, 11, 12, 13, 16, 17, 19, 21, 22, 24, 25, 26, 36, 43, 47, 54, 55, 60, 64, 69, 72, 73, 74, 76, 78, 80, 81, 82,
     83, 85, 86, 89, 93, 94, 101, 102},
    {5, 6, 11, 14, 15, 19, 20, 23, 24, 25, 26, 31, 36, 41, 51, 52, 54, 55, 57, 59, 60, 64, 69, 72, 73, 74, 75, 77, 81,
     82, 83, 85, 86, 88, 89, 93, 102},
    {5, 6, 11, 14, 15, 19, 21, 23, 24, 25, 26, 31, 36, 41, 54, 56, 57, 59, 60, 63, 64, 69, 72, 73, 74, 76, 79, 80, 81,
     82, 84, 85, 86, 88, 89, 93, 94},
    {5, 6, 7, 10, 13, 14, 15, 20, 23, 25, 26, 30, 37, 46, 48, 49, 51, 52, 53, 54, 55, 59, 60, 64, 72, 76, 81, 82, 85,
     86, 87, 88, 89, 97, 98, 104},
    {5, 9, 12, 13, 14, 15, 19, 20, 25, 26, 31, 33, 36, 45, 46, 51, 52, 53, 55, 57, 59, 60, 63, 64, 69, 72, 75, 82, 83,
     87, 88, 89, 93, 98, 101},
    {3, 4, 6, 7, 12, 13, 16, 17, 20, 22, 24, 26, 33, 36, 39, 48, 51, 54, 55, 57, 60, 64, 68, 69, 70, 72, 73, 74, 81, 82,
     87, 89, 93, 94, 97},
    {1, 5, 10, 13, 14, 15, 20, 22, 30, 36, 37, 39, 48, 51, 52, 54, 55, 58, 59, 64, 71, 72, 75, 82, 85, 86, 87, 88, 89,
     92, 97, 98, 100},
    {5, 7, 13, 14, 15, 20, 25, 26, 30, 33, 36, 42, 47, 51, 52, 54, 55, 58, 59, 64, 69, 70, 71, 72, 75, 82, 87, 88, 89,
     91, 92, 97, 98},
    {3, 4, 6, 16, 17, 19, 20, 23, 24, 26, 33, 35, 36, 39, 41, 43, 51, 52, 54, 55, 58, 60, 63, 64, 72, 75, 81, 83, 89,
     92, 99},
    {12, 13, 15, 23, 25, 26, 30, 31, 33, 36, 37, 42, 44, 46, 49, 54, 56, 64, 69, 70, 71, 72, 82, 87, 88, 89, 91, 92, 97,
     98},
    {6, 7, 13, 14, 15, 17, 20, 24, 30, 32, 37, 39, 41, 43, 50, 51, 52, 54, 55, 58, 60, 64, 70, 75, 82, 86, 87, 97, 100,
     102},
    {6, 12, 13, 15, 17, 18, 30, 34, 37, 46, 48, 50, 51, 52, 53, 55, 58, 59, 61, 64, 66, 68, 82, 86, 87, 99, 100, 101,
     102},
    {3, 4, 6, 13, 16, 17, 22, 26, 33, 36, 39, 40, 48, 53, 60, 64, 68, 69, 70, 72, 73, 74, 81, 82, 87, 89, 93, 97, 101},
    {3, 4, 6, 8, 12, 13, 16, 17, 19, 21, 22, 29, 35, 39, 40, 47, 51, 54, 57, 60, 62, 64, 71, 81, 89, 92, 96, 101, 103},
    {2, 6, 10, 14, 15, 16, 17, 19, 22, 25, 26, 36, 50, 53, 55, 59, 63, 64, 72, 73, 74, 81, 82, 87, 89, 93, 95, 98},
    {3, 4, 5, 9, 12, 13, 19, 20, 23, 24, 26, 35, 39, 51, 52, 53, 56, 58, 60, 62, 64, 71, 75, 88, 89, 90, 92, 103},
    {6, 15, 16, 17, 37, 38, 39, 50, 51, 53, 54, 55, 58, 59, 60, 64, 69, 75, 81, 86, 88, 100, 102},
    {3, 4, 31, 35, 48, 49, 51, 52, 58, 62, 64, 69, 71, 82, 89, 90, 92, 95, 99, 100, 103, 106},
    {3, 4, 12, 13, 23, 26, 35, 39, 42, 51, 54, 55, 57, 62, 64, 71, 87, 89, 92, 103, 107},
    {3, 4, 6, 17, 19, 26, 29, 31, 33, 35, 47, 62, 63, 64, 71, 87, 89, 92, 103, 105, 106},
    {7, 12, 13, 14, 15, 21, 25, 26, 35, 39, 51, 52, 56, 58, 65, 87, 89, 92, 97, 100},
    {3, 4, 26, 28, 51, 52, 53, 54, 55, 58, 59, 65, 99, 100, 106, 107},
    {3, 4, 26, 27, 28, 41, 51, 52, 54, 55, 59, 60, 65, 106, 107},
    {27, 28, 39, 40, 42, 51, 52, 53, 54, 55, 57, 65, 108},
    {21, 27, 28, 53, 54, 56, 58, 59, 65, 88, 109},
]

# === 工具函数（使用配置中的范围） ===
def get_bounds():
    mins = np.array([EXPERIMENT_CONFIG['MIN_X'], EXPERIMENT_CONFIG['MIN_Y'], EXPERIMENT_CONFIG['MIN_Z']])
    maxs = np.array([EXPERIMENT_CONFIG['MAX_X'], EXPERIMENT_CONFIG['MAX_Y'], EXPERIMENT_CONFIG['MAX_Z']])
    return mins, maxs

def clip_state(state):
    mins, maxs = get_bounds()
    return np.clip(state, mins, maxs)

def denormalize_state(normalized_state):
    mins, maxs = get_bounds()
    return normalized_state * (maxs - mins) / 2 + (mins + maxs) / 2

def coverage_similarity(triggered, target_path):
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0
    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)

def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']
    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']
    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']
    return reward

# ==================== section6 触发函数（顶层定义，直接接受 x,y,z） ====================
def section6_thermal_electrical_flow_hybrid(x, y, z):
    triggered = set()

    if (25 < x < 75) != (25 < x < 80):
        triggered.add(1)
    if (25 < x < 75) != (30 < x < 75):
        triggered.add(2)

    if (25 < y < 75) != (40 < y < 75):
        triggered.add(3)
    if (25 < y < 75) != (50 < y < 75):
        triggered.add(4)

    if (20 < z < 50) != (30 < z < 50):
        triggered.add(5)
    if (20 < z < 50) != (20 < z < 30):
        triggered.add(6)

    if (x > 40 and x < 60 and y > 35 and y < 65) != (x > 40 and x < 75 and y > 35 and y < 65):
        triggered.add(7)
    if (x > 40 and x < 60 and y > 35 and y < 65) != (x > 40 and x < 60 and y > 20 and y < 65):
        triggered.add(8)

    if (x > 45 and x < 55 and z > 25 and z < 35) != (x > 45 and x < 55 and z > 10 and z < 35):
        triggered.add(9)

    if (y > 40 and y < 60 and z > 25 and z < 35) != (y > 40 and y < 85 and z > 25 and z < 35):
        triggered.add(10)
    if (y > 40 and y < 60 and z > 25 and z < 35) != (y > 40 and y < 60 and z > 30 and z < 40):
        triggered.add(11)

    if (x > 30 and x < 52) != (x > 30 and x < 70):
        triggered.add(12)
    if (x > 30 and x < 52) != (x > 30 and x < 80):
        triggered.add(13)

    if (y > 30 and y < 52) != (y > 30 and y < 70):
        triggered.add(14)
    if (y > 30 and y < 52) != (y > 30 and y < 80):
        triggered.add(15)

    if (z > 20 and z < 32) != (z > 20 and z < 40):
        triggered.add(16)
    if (z > 20 and z < 32) != (z > 20 and z < 50):
        triggered.add(17)

    if (abs(x - 50) < 5) != (abs(x - 50) < 10):
        triggered.add(18)
    if (abs(x - 50) < 5) != (abs(x - 30) < 5):
        triggered.add(19)

    if (abs(y - 50) < 8) != (abs(y - 50) < 15):
        triggered.add(20)
    if (abs(y - 50) < 8) != (abs(y - 20) < 8):
        triggered.add(21)

    if (abs(z - 30) < 3) != (abs(z - 30) < 6):
        triggered.add(22)
    if (abs(z - 30) < 3) != (abs(z - 20) < 3):
        triggered.add(23)

    if (x > 35 and y > 35 and z > 20) != (x > 35 and y > 60 and z > 20):
        triggered.add(24)

    if (x < 75 and y < 75 and z < 45) != (x < 75 and y < 40 and z < 45):
        triggered.add(25)
    if (x < 75 and y < 75 and z < 45) != (x < 75 and y < 30 and z < 45):
        triggered.add(26)

    if (x + y + z > 20 and x + y + z < 180) != (x + y + z > 50 and x + y + z < 180):
        triggered.add(27)
    if (x + y + z > 20 and x + y + z < 180) != (x + y + z > 60 and x + y + z < 180):
        triggered.add(28)

    if (x * y > 1000 and x * y < 4000) != (x * y > 1500 and x * y < 4000):
        triggered.add(29)
    if (x * y > 1000 and x * y < 4000) != (x * y > 1000 and x * y < 5000):
        triggered.add(30)

    if (x * z > 1200 and x * z < 3400) != (x * z > 1500 and x * z < 3400):
        triggered.add(31)
    if (x * z > 1200 and x * z < 3400) != (x * z > 1200 and x * z < 3000):
        triggered.add(32)

    if (y * z > 1000 and y * z < 3000) != (y * z > 1500 and y * z < 3000):
        triggered.add(33)
    if (y * z > 1000 and y * z < 3000) != (y * z > 1000 and y * z < 3500):
        triggered.add(34)

    if ((x + y + z) / 3 > 35 and (x + y + z) / 3 < 55) != ((x + y + z) / 3 > 40 and (x + y + z) / 3 < 55):
        triggered.add(35)
    if ((x + y + z) / 3 > 35 and (x + y + z) / 3 < 55) != ((x + y + z) / 3 > 35 and (x + y + z) / 3 < 40):
        triggered.add(36)

    if (math.sqrt(x ** 2 + y ** 2 + z ** 2) > 180) != (math.sqrt(x ** 2 + y ** 2 + z ** 2) > 100):
        triggered.add(37)
    if (math.sqrt(x ** 2 + y ** 2 + z ** 2) > 180) != (math.sqrt(x ** 2 + y ** 2 + z ** 2) > 120):
        triggered.add(38)

    if (x / y > 1 and x / y < 3) != (x / y > 1 and x / y < 1.2):
        triggered.add(39)
    if (x / y > 1 and x / y < 3) != (x / y > 1 and x / y < 2):
        triggered.add(40)

    if (x / z > 1.5 and x / z < 4.5) != (x / z > 1.8 and x / z < 4.5):
        triggered.add(41)
    if (x / z > 1.5 and x / z < 4.5) != (x / z > 1.5 and x / z < 3):
        triggered.add(42)

    if (y / z > 1.2 and y / z < 4.0) != (y / z > 1.4 and y / z < 4.0):
        triggered.add(43)
    if (y / z > 1.2 and y / z < 4.0) != (y / z > 1.2 and y / z < 3.5):
        triggered.add(44)

    if ((x - 50) + (y - 50) > 10 and (x - 50) - (y - 50) < 10) != ((x - 50) + (y - 50) > 20 and (x - 50) - (y - 50) < 10):
        triggered.add(45)
    if ((x - 50) + (y - 50) > 10 and (x - 50) - (y - 50) < 10) != ((x - 50) + (y - 50) > 50 and (x - 50) - (y - 50) < 10):
        triggered.add(46)

    if ((x - 50) + (z - 30) * 2 > 15 and (x - 50) + (z - 30) * 2 < 15) != ((x - 50) + (z - 30) * 2 > 10 and (x - 50) + (z - 30) * 2 < 15):
        triggered.add(47)
    if ((x - 50) + (z - 30) * 2 > 15 and (x - 50) + (z - 30) * 2 < 15) != ((x - 50) + (z - 30) * 2 > 15 and (x - 50) + (z - 30) * 2 < 40):
        triggered.add(48)

    if ((y - 50) + (z - 30) > 20 and (y - 50) + (z - 30) < 20) != ((y - 50) + (z - 30) > 10 and (y - 50) + (z - 30) < 20):
        triggered.add(49)
    if ((y - 50) + (z - 30) > 20 and (y - 50) + (z - 30) < 20) != ((y - 50) + (z - 30) > 20 and (y - 50) + (z - 30) < 40):
        triggered.add(50)

    if (abs((x - 50) - (y - 50)) < 8) != (abs((x - 50) - (y - 50)) < 30):
        triggered.add(51)
    if (abs((x - 50) - (y - 50)) < 8) != (abs((x - 45) - (y - 50)) < 20):
        triggered.add(52)

    if (abs((x - 50) - (z - 30) * 2) < 10) != (abs((x - 50) - (z - 30) * 2) < 30):
        triggered.add(53)
    if (abs((x - 50) - (z - 30) * 2) < 10) != (abs((x - 80) - (z - 30) * 2) < 10):
        triggered.add(54)

    if (abs((y - 50) - (z - 30)) < 5) != (abs((y - 50) - (z - 30)) < 20):
        triggered.add(55)
    if (abs((y - 50) - (z - 30)) < 5) != (abs((y - 80) - (z - 30)) < 5):
        triggered.add(56)

    if (x / (y + 10) > 0.7 and x / (y + 10) < 1.1) != (x / (y + 10) > 0.8 and x / (y + 10) < 1.5):
        triggered.add(57)
    if (x / (y + 10) > 0.7 and x / (y + 10) < 1.1) != (x / (y + 10) > 0.5 and x / (y + 10) < 1.0):
        triggered.add(58)

    if (y / (z + 20) > 1.0 and y / (z + 20) < 1.8) != (y / (z + 20) > 1.5 and y / (z + 20) < 1.8):
        triggered.add(59)

    if (z / (x / 2) > 0.8 and z / (x / 2) < 1.4) != (z / (x / 2) > 1.5 and z / (x / 2) < 1.4):
        triggered.add(60)
    if (z / (x / 2) > 0.8 and z / (x / 2) < 1.4) != (z / (x / 2) > 0.8 and z / (x / 2) < 1.9):
        triggered.add(61)

    if (x * y * z > 50000 and x * y * z < 150000) != (x * y * z > 30000 and x * y * z < 150000):
        triggered.add(62)

    if (x * 0.4 + y * 0.4 + z * 0.8 > 35 and x * 0.4 + y * 0.4 + z * 0.8 < 65) != (x * 0.4 + y * 0.4 + z * 0.8 > 35 and x * 0.4 + y * 0.4 + z * 0.8 < 60):
        triggered.add(63)

    if ((x * y * z) ** (1 / 3) > 25) != ((x * y * z) ** (1 / 4) > 25):
        triggered.add(64)
    if ((x * y * z) ** (1 / 3) > 25) != ((x * y * z) ** (1 / 2) > 25):
        triggered.add(65)

    if ((x - 50) * (y - 50) > 200 and (x - 50) * (y - 50) < 200) != ((x - 50) * (y - 50) > 150 and (x - 50) * (y - 50) < 200):
        triggered.add(66)

    if ((x - 50) * (z - 30) > 150 and (x - 50) * (z - 30) < 150) != ((x - 50) * (z - 30) > 100 and (x - 50) * (z - 30) < 150):
        triggered.add(67)
    if ((x - 50) * (z - 30) > 150 and (x - 50) * (z - 30) < 150) != ((x - 30) * (z - 30) > 150 and (x - 50) * (z - 30) < 150):
        triggered.add(68)

    if (abs(x / 50 + y / 50 + z / 30 - 4) < 0.5) != (abs(x / 50 + y / 50 + z / 30 - 4) < 1):
        triggered.add(69)
    if (abs(x / 50 + y / 50 + z / 30 - 4) < 0.5) != (abs(x / 40 + y / 50 + z / 30 - 4) < 0.5):
        triggered.add(70)

    if (abs((x / 50) * (y / 50) * (z / 30) - 1) < 0.3) != (abs((x / 50) * (y / 50) * (z / 30) - 1) < 0.8):
        triggered.add(71)
    if (abs((x / 50) * (y / 50) * (z / 30) - 1) < 0.3) != (abs((x / 80) * (y / 50) * (z / 30) - 1) < 0.3):
        triggered.add(72)

    if ((x + y) / 2 > 45 and (x + y) / 2 < 55 and z > 28 and z < 69) != ((x + y) / 2 > 55 and (x + y) / 2 < 55 and z > 38 and z < 69):
        triggered.add(73)
    if ((x + y) / 2 > 45 and (x + y) / 2 < 55 and z > 28 and z < 69) != ((x + y) / 2 > 55 and (x + y) / 2 < 53 and z > 48 and z < 69):
        triggered.add(74)

    if (abs(x - y) < 10 and z > 20 and z < 60) != (abs(x - y) < 20 and z > 20 and z < 60):
        triggered.add(75)
    if (abs(x - y) < 10 and z > 20 and z < 60) != (abs(x - y) < 10 and z > 40 and z < 60):
        triggered.add(76)

    if (math.sqrt((x - 50) ** 2 + (y - 50) ** 2) < 8 and abs(z - 30) < 3) != (math.sqrt((x - 50) ** 2 + (y - 50) ** 2) < 10 and abs(z - 30) < 3):
        triggered.add(77)
    if (math.sqrt((x - 50) ** 2 + (y - 50) ** 2) < 8 and abs(z - 30) < 3) != (math.sqrt((x - 50) ** 2 + (y - 50) ** 2) < 8 and abs(z - 30) < 10):
        triggered.add(78)

    if (x / 50 > 0.9 and x / 50 < 1.1 and y / 50 > 0.9 and y / 50 < 1.1) != (x / 50 > 1 and x / 50 < 1.1 and y / 50 > 0.9 and y / 50 < 1.1):
        triggered.add(79)
    if (x / 50 > 0.9 and x / 50 < 1.1 and y / 50 > 0.9 and y / 50 < 1.1) != (x / 50 > 1.2 and x / 50 < 1.1 and y / 50 > 0.9 and y / 50 < 1.1):
        triggered.add(80)

    if (z / 30 > 0.85 and z / 30 < 1) != (z / 30 > 0.85 and z / 30 < 1.3):
        triggered.add(81)

    if ((x / 50 + y / 50 + z / 30) / 3 > 0.9 and (x / 50 + y / 50 + z / 30) / 3 < 1) != ((x / 50 + y / 50 + z / 30) / 3 > 0.9 and (x / 50 + y / 50 + z / 30) / 3 < 1.5):
        triggered.add(82)

    if (max(abs(x / 50 - 1), abs(y / 50 - 1), abs(z / 30 - 1)) < 0.15) != (max(abs(x / 50 - 1), abs(y / 50 - 1), abs(z / 30 - 1)) < 0.25):
        triggered.add(83)
    if (max(abs(x / 50 - 1), abs(y / 50 - 1), abs(z / 30 - 1)) < 0.15) != (max(abs(x / 25 - 1), abs(y / 50 - 1), abs(z / 30 - 1)) < 0.15):
        triggered.add(84)

    if (min(x / 50, y / 50, z / 30) > 0.8) != (min(x / 50, y / 50, z / 30) > 1):
        triggered.add(85)
    if (min(x / 50, y / 50, z / 30) > 0.8) != (min(x / 50, y / 50, z / 30) > 1.8):
        triggered.add(86)

    if (max(x / 50, y / 50, z / 30) < 1.2) != (max(x / 50, y / 50, z / 30) < 1.6):
        triggered.add(87)

    if (abs(max(x, y, z * 2) - min(x, y, z * 2)) < 20) != (abs(max(x, y, z * 3) - min(x, y, z * 3)) < 20):
        triggered.add(88)

    if (x + y + z * 1.5 > 100 and x + y + z * 1.5 < 300) != (x + y + z * 1.5 > 180 and x + y + z * 1.5 < 300):
        triggered.add(89)

    if (x * y * z > 40000 and x * y * z < 120000) != (x * y * z > 45000 and x * y * z < 120000):
        triggered.add(90)
    if (x * y * z > 40000 and x * y * z < 120000) != (x * y * z > 40000 and x * y * z < 95000):
        triggered.add(91)

    if (abs((x + y + z) / 3 - 45) < 3) != (abs((x + y + z) / 3 - 45) < 10):
        triggered.add(92)
    if (abs((x + y + z) / 3 - 45) < 3) != (abs((x + y + z) / 3 - 30) < 3):
        triggered.add(93)

    if (abs(x * y / z - 75) < 5) != (abs(x * y / z - 75) < 10):
        triggered.add(94)

    if (x < 30 or y < 30 or z < 15) != (x < 20 or y < 30 or z < 15):
        triggered.add(95)
    if (x < 30 or y < 30 or z < 15) != (x < 30 or y < 20 or z < 15):
        triggered.add(96)

    if (x > 75 or y > 75 or z > 45) != (x > 60 or y > 75 or z > 45):
        triggered.add(97)
    if (x > 75 or y > 75 or z > 45) != (x > 75 or y > 60 or z > 45):
        triggered.add(98)

    if (abs(x / y - 1.0) > 0.25) != (abs(x / y - 1.0) > 0.3):
        triggered.add(99)
    if (abs(x / y - 1.0) > 0.25) != (abs(x / y - 0.95) > 0.25):
        triggered.add(100)

    if (abs(x / z - 1.8) > 0.3) != (abs(x / z - 1.8) > 0.4):
        triggered.add(101)

    if (abs(y / z - 1.7) > 0.3) != (abs(y / z - 1.6) > 0.3):
        triggered.add(102)

    if (x + y + z < 110 or x + y + z > 190) != (x + y + z < 115 or x + y + z > 190):
        triggered.add(103)

    if (x * y * z < 60000 or x * y * z > 140000) != (x * y * z < 60000 or x * y * z > 135000):
        triggered.add(104)

    if (x < 35 and y < 35) != (x < 30 and y < 35):
        triggered.add(105)
    if (x < 35 and y < 35) != (x < 35 and y < 30):
        triggered.add(106)

    if (y < 35 and z < 20) != (y < 30 and z < 20):
        triggered.add(107)

    if (x < 25 and y < 25 and z < 15) != (x < 22 and y < 25 and z < 15):
        triggered.add(108)
    if (x < 25 and y < 25 and z < 15) != (x < 25 and y < 22 and z < 15):
        triggered.add(109)

    return triggered

# ========== 设置执行函数为 section6 ==========
execute_Tr = section6_thermal_electrical_flow_hybrid

# === DQN 网络 ===
class DQNNetwork(nn.Module):
    def __init__(self, action_size=30):
        super(DQNNetwork, self).__init__()
        hidden_dim = EXPERIMENT_CONFIG['HIDDEN_DIM']
        self.fc1 = nn.Linear(3, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)
        self.output = nn.Linear(hidden_dim, action_size)

    def forward(self, x):
        x = torch.relu(self.fc1(x))
        x = torch.relu(self.fc2(x))
        x = torch.relu(self.fc3(x))
        return self.output(x)

# === PathReplayBuffer ===
class PathReplayBuffer:
    def __init__(self, path_idx, capacity=20000):
        self.path_idx = path_idx
        self.capacity = capacity
        self.buffer = deque(maxlen=capacity)
        self.similarities = deque(maxlen=capacity)

    def push(self, state, action, reward, next_state, done, similarity):
        self.buffer.append((state, action, reward, next_state, done))
        self.similarities.append(similarity)

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return None
        batch = random.sample(self.buffer, batch_size)
        state, action, reward, next_state, done = map(np.stack, zip(*batch))
        return (
            torch.FloatTensor(state).to(device),
            torch.LongTensor(action).to(device),
            torch.FloatTensor(reward).to(device),
            torch.FloatTensor(next_state).to(device),
            torch.BoolTensor(done).to(device)
        )

    def get_top_k(self, k=20):
        if len(self.buffer) == 0:
            return []
        samples_with_sim = list(zip(self.buffer, self.similarities))
        samples_with_sim.sort(key=lambda x: x[1], reverse=True)
        top_k = samples_with_sim[:k]
        results = []
        for (state, _, _, _, _), similarity in top_k:
            original_state = denormalize_state(state)
            original_state_int = np.round(original_state).astype(int)
            results.append({
                'state': original_state_int,
                'similarity': similarity,
                'triggered': execute_Tr(*original_state_int)
            })
        return results

    def __len__(self):
        return len(self.buffer)

# === ImprovedDQNAgent ===
class ImprovedDQNAgent:
    def __init__(self, num_paths, action_size=30):
        self.action_size = action_size
        self.num_paths = num_paths
        self.epsilon = 0.9
        self.epsilon_min = 0.1
        self.epsilon_decay = 0.995

        self.q_network = DQNNetwork(action_size).to(device)
        self.target_network = DQNNetwork(action_size).to(device)
        self.optimizer = optim.Adam(self.q_network.parameters(), lr=EXPERIMENT_CONFIG['LEARNING_RATE'])

        self.replay_buffers = {}
        for path_idx in range(num_paths):
            self.replay_buffers[path_idx] = PathReplayBuffer(path_idx, EXPERIMENT_CONFIG['REPLAY_BUFFER_CAPACITY'])

        self.replay_train_count = 0
        self.update_target_network()

    def discrete_to_action_delta(self, action_idx):
        delta_values = [5, 3, 2, 1, 0.5, -0.5, -1, -2, -3, -5]
        if action_idx >= 30:
            action_idx = action_idx % 30
        dim = action_idx // 10
        delta_idx = action_idx % 10
        delta = delta_values[delta_idx]
        action_delta = np.zeros(3)
        action_delta[dim] = delta
        return action_delta

    def get_action(self, state):
        mins, maxs = get_bounds()
        normalized_state = (state - (mins + maxs) / 2) / ((maxs - mins) / 2)
        if random.random() < self.epsilon:
            action_idx = random.randint(0, self.action_size - 1)
        else:
            state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)
            with torch.no_grad():
                q_values = self.q_network(state_tensor)
                action_idx = q_values.argmax().item()
        action_delta = self.discrete_to_action_delta(action_idx)
        return action_delta, action_idx

    def store_experience(self, path_idx, state, action_idx, reward, next_state, done, similarity):
        mins, maxs = get_bounds()
        normalized_state = (state - (mins + maxs) / 2) / ((maxs - mins) / 2)
        normalized_next_state = (next_state - (mins + maxs) / 2) / ((maxs - mins) / 2)
        self.replay_buffers[path_idx].push(
            normalized_state, action_idx, reward,
            normalized_next_state, done, similarity
        )

    def replay_train(self, path_idx):
        batch = self.replay_buffers[path_idx].sample(EXPERIMENT_CONFIG['REPLAY_BATCH_SIZE'])
        if batch is None:
            return
        states, actions, rewards, next_states, dones = batch
        current_q_values = self.q_network(states).gather(1, actions.unsqueeze(1))
        with torch.no_grad():
            next_q_values = self.target_network(next_states).max(1)[0]
            target_q_values = rewards + (0.99 * next_q_values * ~dones)
        loss = F.mse_loss(current_q_values.squeeze(), target_q_values)
        self.optimizer.zero_grad()
        loss.backward()
        torch.nn.utils.clip_grad_norm_(self.q_network.parameters(), 1.0)
        self.optimizer.step()
        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay
        self.replay_train_count += 1
        if self.replay_train_count % 2 == 0:
            self.update_target_network()

    def update_target_network(self):
        self.target_network.load_state_dict(self.q_network.state_dict())

    def get_all_top_k(self, k=20):
        results = {}
        for path_idx in range(self.num_paths):
            results[path_idx] = self.replay_buffers[path_idx].get_top_k(k)
        return results

    def get_buffer_stats(self):
        stats = {}
        for path_idx in range(self.num_paths):
            stats[path_idx] = len(self.replay_buffers[path_idx])
        return stats

# === 性能计算 ===
def calculate_run_performance(run_idx, dqn_results, training_time, total_steps, update_count, agent):
    target_paths = targetPaths
    num_paths = len(target_paths)
    all_similarities = []
    total_reward = 0
    total_samples = 0
    all_rewards = []
    for path_idx in range(num_paths):
        samples = dqn_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']
            total_reward += reward
            all_rewards.append(reward)
            all_similarities.append(similarity)
            total_samples += 1
    average_reward = total_reward / total_samples if total_samples > 0 else 0
    convergence = np.mean(all_similarities) if all_similarities else 0
    environment_adaptability = 1 / (np.std(all_similarities) + 1e-8) if len(all_similarities) > 1 else 0
    generalization_ability = convergence
    computational_efficiency = total_steps / training_time if training_time > 0 else 0
    policy_update_frequency = update_count / training_time if training_time > 0 else 0
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0
    return {
        'Run': run_idx + 1,
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }

# === Excel导出 ===
def export_to_excel(all_dqn_results, all_performance_data, target_paths, output_path="DQN_20_run.xlsx"):
    print("\n导出Excel...")
    all_dqn_summary_data = []
    all_dqn_detailed_data = []
    for run_idx, (dqn_results, performance_data) in enumerate(zip(all_dqn_results, all_performance_data)):
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = dqn_results[path_idx]
            if len(samples) == 0:
                all_dqn_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue
            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = 'Yes' if perfect_count > 0 else 'No'
            all_dqn_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })
            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']
                all_dqn_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Sample ID': sample_idx + 1,
                    'X': int(state[0]),
                    'Y': int(state[1]),
                    'Z': int(state[2]),
                    'Similarity': round(similarity, 4),
                    'Perfect Coverage': 'Yes' if abs(similarity - 1.0) < 0.001 else 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path))),
                    'Triggered Rules': ', '.join(map(str, sorted(triggered))),
                    'Hit Rule Count': len(target_path.intersection(triggered)),
                    'Target Rule Count': len(target_path)
                })
    dqn_summary_df = pd.DataFrame(all_dqn_summary_data)
    dqn_detailed_df = pd.DataFrame(all_dqn_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        dqn_summary_df.to_excel(writer, sheet_name='DQNPath', index=False)
        dqn_detailed_df.to_excel(writer, sheet_name='DQNDetailed Sample Data', index=False)
        selected_columns = [
            'Run', 'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df[selected_columns].to_excel(writer, sheet_name='Metric', index=False)
        workbook = writer.book
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        for sheet_name in ['DQNPath', 'DQNDetailed Sample Data', 'Metric']:
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
    print(f"Excel已保存: {output_path}")

# === 训练流程 ===
def train_dqn_workflow():
    print("=" * 80)
    print("开始DQN训练")
    print(f"路径数: {len(targetPaths)}, 每路径样本数: {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']}")
    print("=" * 80)
    num_paths = len(targetPaths)
    agent = ImprovedDQNAgent(num_paths=num_paths)
    start_time = time.time()
    total_steps = 0

    # 生成初始样本
    path_samples = {}
    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            state = np.array([
                random.randint(EXPERIMENT_CONFIG['MIN_X'], EXPERIMENT_CONFIG['MAX_X']),
                random.randint(EXPERIMENT_CONFIG['MIN_Y'], EXPERIMENT_CONFIG['MAX_Y']),
                random.randint(EXPERIMENT_CONFIG['MIN_Z'], EXPERIMENT_CONFIG['MAX_Z'])
            ], dtype=np.float32)
            samples.append(state)
        path_samples[path_idx] = samples

    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size
    num_rounds = EXPERIMENT_CONFIG['NUM_ROUNDS']

    for path_idx in range(num_paths):
        target_path = targetPaths[path_idx]
        print(f"\n开始训练路径 {path_idx+1}/{num_paths}, 目标规则数: {len(target_path)}")
        for round_idx in range(num_rounds):
            print(f"  轮次 {round_idx+1}/{num_rounds}")
            for batch_idx in range(num_batches):
                batch_samples = path_samples[path_idx][batch_idx*batch_size:(batch_idx+1)*batch_size]
                batch_rewards = []
                batch_similarities = []
                for initial_state in batch_samples:
                    state = initial_state.copy()
                    episode_reward = 0
                    final_similarity = 0
                    for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                        action_delta, action_idx = agent.get_action(state)
                        next_state = state + action_delta
                        next_state = clip_state(next_state)
                        triggered = execute_Tr(*next_state)
                        reward = unified_reward_function(triggered, target_path)
                        similarity = coverage_similarity(triggered, target_path)
                        done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)
                        agent.store_experience(path_idx, state, action_idx, reward, next_state, done, similarity)
                        state = next_state
                        episode_reward += reward
                        final_similarity = similarity
                        total_steps += 1
                    batch_rewards.append(episode_reward)
                    batch_similarities.append(final_similarity)
                avg_reward = np.mean(batch_rewards)
                avg_sim = np.mean(batch_similarities)
                print(f"    批次 {batch_idx+1}/{num_batches}: 平均奖励={avg_reward:.2f}, 平均相似度={avg_sim:.4f}, epsilon={agent.epsilon:.3f}")
                agent.replay_train(path_idx)
    training_time = time.time() - start_time
    print(f"\n训练完成，总耗时: {training_time:.2f}秒，总步数: {total_steps}")
    dqn_top_k_results = agent.get_all_top_k(EXPERIMENT_CONFIG['TOP_K_SAMPLES'])
    return agent, dqn_top_k_results, training_time, total_steps, agent.replay_train_count

# === 主程序 ===
def main():
    print("=" * 80)
    print("DQN - 20次运行实验")
    print(f"状态空间范围: X[1,100], Y[1,100], Z[1,60]")
    print("=" * 80)
    all_dqn_results = []
    all_performance_data = []
    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n开始第 {run_idx+1}/{EXPERIMENT_CONFIG['NUM_RUNS']} 次运行")
        dqn_agent, dqn_results, training_time, total_steps, update_count = train_dqn_workflow()
        performance_data = calculate_run_performance(run_idx, dqn_results, training_time, total_steps, update_count, dqn_agent)
        all_dqn_results.append(dqn_results)
        all_performance_data.append(performance_data)
        print(f"运行 {run_idx+1} 完成: 总奖励={performance_data['Total Reward']}, 收敛度={performance_data['Convergence']}")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"DQN_20_run_{timestamp}.xlsx"
    export_to_excel(all_dqn_results, all_performance_data, targetPaths, output_path)
    print("\n20次运行全部完成！")

if __name__ == "__main__":
    main()