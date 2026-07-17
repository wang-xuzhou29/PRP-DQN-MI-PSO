import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
from torch.distributions import Normal
import numpy as np
import random
import time
import math
from collections import deque
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# 设备配置
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# === 目标路径组 ===
targetPaths = [
    {4, 6, 11, 12, 13, 16, 17, 18, 21, 24, 25, 26, 36, 39, 41, 43, 48, 51, 52, 54, 55, 58, 60, 64, 68, 69, 70, 71, 72,
     73, 74, 75, 81, 82, 83, 85, 86, 87, 89, 91, 93, 99, 102},
    {6, 10, 12, 13, 14, 15, 16, 17, 18, 20, 22, 25, 26, 36, 41, 45, 46, 47, 49, 51, 52, 54, 55, 57, 59, 60, 64, 69, 70,
     71, 72, 76, 81, 82, 86, 87, 88, 89, 91, 92, 98, 102},
    {6, 7, 11, 12, 13, 14, 15, 16, 17, 21, 24, 25, 26, 36, 39, 41, 43, 48, 49, 51, 52, 54, 55, 58, 60, 64, 67, 68, 72,
     75, 82, 86, 87, 89, 92, 97, 100, 102, 104},
    {6, 11, 12, 13, 14, 15, 16, 17, 19, 21, 24, 25, 26, 36, 48, 49, 53, 56, 57, 61, 64, 68, 69, 70, 71, 72, 73, 74, 76,
     78, 81, 82, 83, 86, 87, 89, 91, 92, 101},
    {6, 7, 13, 14, 15, 16, 17, 20, 22, 24, 25, 26, 30, 36, 39, 48, 49, 51, 52, 53, 56, 58, 59, 60, 64, 66, 68, 72, 75,
     81, 82, 86, 87, 88, 89, 92, 97, 100, 104},
    {4, 6, 12, 13, 16, 17, 19, 21, 22, 24, 25, 26, 33, 36, 39, 41, 43, 47, 51, 52, 54, 55, 60, 64, 69, 72, 73, 74, 76,
     78, 81, 82, 84, 85, 86, 89, 93, 94, 100},
    {5, 7, 11, 12, 13, 14, 15, 21, 22, 24, 25, 26, 33, 36, 39, 47, 51, 52, 53, 54, 55, 58, 59, 64, 69, 70, 71, 72, 75,
     82, 85, 86, 87, 88, 89, 91, 92, 97, 100},
    {6, 10, 14, 15, 16, 17, 18, 20, 22, 24, 25, 26, 31, 36, 49, 51, 52, 53, 55, 58, 59, 61, 64, 69, 73, 74, 75, 81, 82,
     83, 85, 86, 87, 89, 93, 94, 99, 100},
    {5, 7, 11, 13, 14, 15, 21, 23, 24, 25, 26, 30, 36, 39, 48, 51, 53, 54, 55, 58, 59, 60, 64, 66, 71, 72, 75, 82, 85,
     86, 87, 88, 89, 91, 92, 97, 99, 102},
    {6, 10, 12, 13, 15, 16, 17, 19, 23, 25, 26, 36, 41, 46, 50, 51, 52, 54, 55, 58, 59, 60, 64, 71, 72, 75, 81, 82, 86,
     87, 88, 89, 91, 92, 98, 99, 100},
    {4, 6, 11, 12, 13, 16, 17, 19, 21, 22, 24, 25, 26, 36, 43, 47, 54, 55, 60, 64, 69, 72, 73, 74, 76, 78, 80, 81, 82,
     83, 85, 86, 89, 93, 94, 101, 102},
    {5, 6, 11, 14, 15, 19, 20, 23, 24, 25, 26, 31, 36, 41, 51, 52, 54, 55, 57, 59, 60, 64, 69, 72, 73, 74, 75, 77, 81,
     82, 83, 85, 86, 88, 89, 93, 102},
    {5, 6, 11, 14, 15, 19, 21, 23, 24, 25, 26, 31, 36, 41, 54, 56, 57, 59, 60, 63, 64, 69, 72, 73, 74, 76, 79, 80, 81,
     82, 84, 85, 86, 88, 89, 93, 94},
    {5, 6, 7, 10, 13, 14, 15, 20, 23, 25, 26, 30, 37, 46, 48, 49, 51, 52, 53, 54, 55, 59, 60, 64, 72, 76, 81, 82, 85,
     86, 87, 88, 89, 97, 98, 104},
    {5, 9, 12, 13, 14, 15, 19, 20, 25, 26, 31, 33, 36, 45, 46, 51, 52, 53, 55, 57, 59, 60, 63, 64, 69, 72, 75, 82, 83,
     87, 88, 89, 93, 98, 101},
    {3, 4, 6, 7, 12, 13, 16, 17, 20, 22, 24, 26, 33, 36, 39, 48, 51, 54, 55, 57, 60, 64, 68, 69, 70, 72, 73, 74, 81, 82,
     87, 89, 93, 94, 97},
    {1, 5, 10, 13, 14, 15, 20, 22, 30, 36, 37, 39, 48, 51, 52, 54, 55, 58, 59, 64, 71, 72, 75, 82, 85, 86, 87, 88, 89,
     92, 97, 98, 100},
    {5, 7, 13, 14, 15, 20, 25, 26, 30, 33, 36, 42, 47, 51, 52, 54, 55, 58, 59, 64, 69, 70, 71, 72, 75, 82, 87, 88, 89,
     91, 92, 97, 98},
    {3, 4, 6, 16, 17, 19, 20, 23, 24, 26, 33, 35, 36, 39, 41, 43, 51, 52, 54, 55, 58, 60, 63, 64, 72, 75, 81, 83, 89,
     92, 99},
    {12, 13, 15, 23, 25, 26, 30, 31, 33, 36, 37, 42, 44, 46, 49, 54, 56, 64, 69, 70, 71, 72, 82, 87, 88, 89, 91, 92, 97,
     98},
    {6, 7, 13, 14, 15, 17, 20, 24, 30, 32, 37, 39, 41, 43, 50, 51, 52, 54, 55, 58, 60, 64, 70, 75, 82, 86, 87, 97, 100,
     102},
    {6, 12, 13, 15, 17, 18, 30, 34, 37, 46, 48, 50, 51, 52, 53, 55, 58, 59, 61, 64, 66, 68, 82, 86, 87, 99, 100, 101,
     102},
    {3, 4, 6, 13, 16, 17, 22, 26, 33, 36, 39, 40, 48, 53, 60, 64, 68, 69, 70, 72, 73, 74, 81, 82, 87, 89, 93, 97, 101},
    {3, 4, 6, 8, 12, 13, 16, 17, 19, 21, 22, 29, 35, 39, 40, 47, 51, 54, 57, 60, 62, 64, 71, 81, 89, 92, 96, 101, 103},
    {2, 6, 10, 14, 15, 16, 17, 19, 22, 25, 26, 36, 50, 53, 55, 59, 63, 64, 72, 73, 74, 81, 82, 87, 89, 93, 95, 98},
    {3, 4, 5, 9, 12, 13, 19, 20, 23, 24, 26, 35, 39, 51, 52, 53, 56, 58, 60, 62, 64, 71, 75, 88, 89, 90, 92, 103},
    {6, 15, 16, 17, 37, 38, 39, 50, 51, 53, 54, 55, 58, 59, 60, 64, 69, 75, 81, 86, 88, 100, 102},
    {3, 4, 31, 35, 48, 49, 51, 52, 58, 62, 64, 69, 71, 82, 89, 90, 92, 95, 99, 100, 103, 106},
    {3, 4, 12, 13, 23, 26, 35, 39, 42, 51, 54, 55, 57, 62, 64, 71, 87, 89, 92, 103, 107},
    {3, 4, 6, 17, 19, 26, 29, 31, 33, 35, 47, 62, 63, 64, 71, 87, 89, 92, 103, 105, 106},
    {7, 12, 13, 14, 15, 21, 25, 26, 35, 39, 51, 52, 56, 58, 65, 87, 89, 92, 97, 100},
    {3, 4, 26, 28, 51, 52, 53, 54, 55, 58, 59, 65, 99, 100, 106, 107},
    {3, 4, 26, 27, 28, 41, 51, 52, 54, 55, 59, 60, 65, 106, 107},
    {27, 28, 39, 40, 42, 51, 52, 53, 54, 55, 57, 65, 108},
    {21, 27, 28, 53, 54, 56, 58, 59, 65, 88, 109},
]

# === 实验配置（X温度1~100, Y电压1~100, Z流量1~60） ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'ACTION_DIM': 3,
    'MIN_VALUES': np.array([1.0, 1.0, 1.0], dtype=np.float32),
    'MAX_VALUES': np.array([100.0, 100.0, 60.0], dtype=np.float32),
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 3,
    'REPLAY_BATCH_SIZE': 64,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'ACTOR_LR': 3e-4,
    'CRITIC_LR': 3e-4,
    'ALPHA_LR': 3e-4,
    'GAMMA': 0.99,
    'TAU': 0.005,
    'NUM_RUNS': 20,
    'TOP_K_SAMPLES': 20,
    'TARGET_PATHS': targetPaths
}


def section6_thermal_electrical_flow_hybrid(x, y, z):
    """第6类: 热电流量混合控制 (温度x, 电压y, 流量z) - 109个有效变异分支"""
    triggered = set()

    # 1-6：单变量边界偏移
    if (25 < x < 75) != (25 < x < 80):
        triggered.add(1)
    if (25 < x < 75) != (30 < x < 75):
        triggered.add(2)
    if (25 < y < 75) != (40 < y < 75):
        triggered.add(3)
    if (25 < y < 75) != (50 < y < 75):
        triggered.add(4)
    if (20 < z < 50) != (30 < z < 50):
        triggered.add(5)
    if (20 < z < 50) != (20 < z < 30):
        triggered.add(6)

    # 7-11：双变量联合区间偏移
    if (x > 40 and x < 60 and y > 35 and y < 65) != (x > 40 and x < 75 and y > 35 and y < 65):
        triggered.add(7)
    if (x > 40 and x < 60 and y > 35 and y < 65) != (x > 40 and x < 60 and y > 20 and y < 65):
        triggered.add(8)
    if (x > 45 and x < 55 and z > 25 and z < 35) != (x > 45 and x < 55 and z > 10 and z < 35):
        triggered.add(9)
    if (y > 40 and y < 60 and z > 25 and z < 35) != (y > 40 and y < 85 and z > 25 and z < 35):
        triggered.add(10)
    if (y > 40 and y < 60 and z > 25 and z < 35) != (y > 40 and y < 60 and z > 30 and z < 40):
        triggered.add(11)

    # 12-17：单变量窄区间扩展
    if (x > 30 and x < 52) != (x > 30 and x < 70):
        triggered.add(12)
    if (x > 30 and x < 52) != (x > 30 and x < 80):
        triggered.add(13)
    if (y > 30 and y < 52) != (y > 30 and y < 70):
        triggered.add(14)
    if (y > 30 and y < 52) != (y > 30 and y < 80):
        triggered.add(15)
    if (z > 20 and z < 32) != (z > 20 and z < 40):
        triggered.add(16)
    if (z > 20 and z < 32) != (z > 20 and z < 50):
        triggered.add(17)

    # 18-23：绝对值区间偏移
    if (abs(x - 50) < 5) != (abs(x - 50) < 10):
        triggered.add(18)
    if (abs(x - 50) < 5) != (abs(x - 30) < 5):
        triggered.add(19)
    if (abs(y - 50) < 8) != (abs(y - 50) < 15):
        triggered.add(20)
    if (abs(y - 50) < 8) != (abs(y - 20) < 8):
        triggered.add(21)
    if (abs(z - 30) < 3) != (abs(z - 30) < 6):
        triggered.add(22)
    if (abs(z - 30) < 3) != (abs(z - 20) < 3):
        triggered.add(23)

    # 24-26：三变量下界/上界偏移
    if (x > 35 and y > 35 and z > 20) != (x > 35 and y > 60 and z > 20):
        triggered.add(24)
    if (x < 75 and y < 75 and z < 45) != (x < 75 and y < 40 and z < 45):
        triggered.add(25)
    if (x < 75 and y < 75 and z < 45) != (x < 75 and y < 30 and z < 45):
        triggered.add(26)

    # 27-34：和/积区间偏移
    if (x + y + z > 20 and x + y + z < 180) != (x + y + z > 50 and x + y + z < 180):
        triggered.add(27)
    if (x + y + z > 20 and x + y + z < 180) != (x + y + z > 60 and x + y + z < 180):
        triggered.add(28)
    if (x * y > 1000 and x * y < 4000) != (x * y > 1500 and x * y < 4000):
        triggered.add(29)
    if (x * y > 1000 and x * y < 4000) != (x * y > 1000 and x * y < 5000):
        triggered.add(30)
    if (x * z > 1200 and x * z < 3400) != (x * z > 1500 and x * z < 3400):
        triggered.add(31)
    if (x * z > 1200 and x * z < 3400) != (x * z > 1200 and x * z < 3000):
        triggered.add(32)
    if (y * z > 1000 and y * z < 3000) != (y * z > 1500 and y * z < 3000):
        triggered.add(33)
    if (y * z > 1000 and y * z < 3000) != (y * z > 1000 and y * z < 3500):
        triggered.add(34)

    # 35-38：均值/模长区间偏移
    if ((x + y + z) / 3 > 35 and (x + y + z) / 3 < 55) != ((x + y + z) / 3 > 40 and (x + y + z) / 3 < 55):
        triggered.add(35)
    if ((x + y + z) / 3 > 35 and (x + y + z) / 3 < 55) != ((x + y + z) / 3 > 35 and (x + y + z) / 3 < 40):
        triggered.add(36)
    if (math.sqrt(x ** 2 + y ** 2 + z ** 2) > 180) != (math.sqrt(x ** 2 + y ** 2 + z ** 2) > 100):
        triggered.add(37)
    if (math.sqrt(x ** 2 + y ** 2 + z ** 2) > 180) != (math.sqrt(x ** 2 + y ** 2 + z ** 2) > 120):
        triggered.add(38)

    # 39-44：比值区间偏移
    if (x / y > 1 and x / y < 3) != (x / y > 1 and x / y < 1.2):
        triggered.add(39)
    if (x / y > 1 and x / y < 3) != (x / y > 1 and x / y < 2):
        triggered.add(40)
    if (x / z > 1.5 and x / z < 4.5) != (x / z > 1.8 and x / z < 4.5):
        triggered.add(41)
    if (x / z > 1.5 and x / z < 4.5) != (x / z > 1.5 and x / z < 3):
        triggered.add(42)
    if (y / z > 1.2 and y / z < 4.0) != (y / z > 1.4 and y / z < 4.0):
        triggered.add(43)
    if (y / z > 1.2 and y / z < 4.0) != (y / z > 1.2 and y / z < 3.5):
        triggered.add(44)

    # 45-50：中心化线性组合偏移
    if ((x - 50) + (y - 50) > 10 and (x - 50) - (y - 50) < 10) != (
            (x - 50) + (y - 50) > 20 and (x - 50) - (y - 50) < 10):
        triggered.add(45)
    if ((x - 50) + (y - 50) > 10 and (x - 50) - (y - 50) < 10) != (
            (x - 50) + (y - 50) > 50 and (x - 50) - (y - 50) < 10):
        triggered.add(46)
    if ((x - 50) + (z - 30) * 2 > 15 and (x - 50) + (z - 30) * 2 < 15) != (
            (x - 50) + (z - 30) * 2 > 10 and (x - 50) + (z - 30) * 2 < 15):
        triggered.add(47)
    if ((x - 50) + (z - 30) * 2 > 15 and (x - 50) + (z - 30) * 2 < 15) != (
            (x - 50) + (z - 30) * 2 > 15 and (x - 50) + (z - 30) * 2 < 40):
        triggered.add(48)
    if ((y - 50) + (z - 30) > 20 and (y - 50) + (z - 30) < 20) != (
            (y - 50) + (z - 30) > 10 and (y - 50) + (z - 30) < 20):
        triggered.add(49)
    if ((y - 50) + (z - 30) > 20 and (y - 50) + (z - 30) < 20) != (
            (y - 50) + (z - 30) > 20 and (y - 50) + (z - 30) < 40):
        triggered.add(50)

    # 51-56：差值绝对值偏移
    if (abs((x - 50) - (y - 50)) < 8) != (abs((x - 50) - (y - 50)) < 30):
        triggered.add(51)
    if (abs((x - 50) - (y - 50)) < 8) != (abs((x - 45) - (y - 50)) < 20):
        triggered.add(52)
    if (abs((x - 50) - (z - 30) * 2) < 10) != (abs((x - 50) - (z - 30) * 2) < 30):
        triggered.add(53)
    if (abs((x - 50) - (z - 30) * 2) < 10) != (abs((x - 80) - (z - 30) * 2) < 10):
        triggered.add(54)
    if (abs((y - 50) - (z - 30)) < 5) != (abs((y - 50) - (z - 30)) < 20):
        triggered.add(55)
    if (abs((y - 50) - (z - 30)) < 5) != (abs((y - 80) - (z - 30)) < 5):
        triggered.add(56)

    # 57-65：归一化比值与幂次偏移
    if (x / (y + 10) > 0.7 and x / (y + 10) < 1.1) != (x / (y + 10) > 0.8 and x / (y + 10) < 1.5):
        triggered.add(57)
    if (x / (y + 10) > 0.7 and x / (y + 10) < 1.1) != (x / (y + 10) > 0.5 and x / (y + 10) < 1.0):
        triggered.add(58)
    if (y / (z + 20) > 1.0 and y / (z + 20) < 1.8) != (y / (z + 20) > 1.5 and y / (z + 20) < 1.8):
        triggered.add(59)
    if (z / (x / 2) > 0.8 and z / (x / 2) < 1.4) != (z / (x / 2) > 1.5 and z / (x / 2) < 1.4):
        triggered.add(60)
    if (z / (x / 2) > 0.8 and z / (x / 2) < 1.4) != (z / (x / 2) > 0.8 and z / (x / 2) < 1.9):
        triggered.add(61)
    if (x * y * z > 50000 and x * y * z < 150000) != (x * y * z > 30000 and x * y * z < 150000):
        triggered.add(62)
    if (x * 0.4 + y * 0.4 + z * 0.8 > 35 and x * 0.4 + y * 0.4 + z * 0.8 < 65) != (
            x * 0.4 + y * 0.4 + z * 0.8 > 35 and x * 0.4 + y * 0.4 + z * 0.8 < 60):
        triggered.add(63)
    if ((x * y * z) ** (1 / 3) > 25) != ((x * y * z) ** (1 / 4) > 25):
        triggered.add(64)
    if ((x * y * z) ** (1 / 3) > 25) != ((x * y * z) ** (1 / 2) > 25):
        triggered.add(65)

    # 66-72：中心化乘积与归一化偏移
    if ((x - 50) * (y - 50) > 200 and (x - 50) * (y - 50) < 200) != (
            (x - 50) * (y - 50) > 150 and (x - 50) * (y - 50) < 200):
        triggered.add(66)
    if ((x - 50) * (z - 30) > 150 and (x - 50) * (z - 30) < 150) != (
            (x - 50) * (z - 30) > 100 and (x - 50) * (z - 30) < 150):
        triggered.add(67)
    if ((x - 50) * (z - 30) > 150 and (x - 50) * (z - 30) < 150) != (
            (x - 30) * (z - 30) > 150 and (x - 50) * (z - 30) < 150):
        triggered.add(68)
    if (abs(x / 50 + y / 50 + z / 30 - 4) < 0.5) != (abs(x / 50 + y / 50 + z / 30 - 4) < 1):
        triggered.add(69)
    if (abs(x / 50 + y / 50 + z / 30 - 4) < 0.5) != (abs(x / 40 + y / 50 + z / 30 - 4) < 0.5):
        triggered.add(70)
    if (abs((x / 50) * (y / 50) * (z / 30) - 1) < 0.3) != (abs((x / 50) * (y / 50) * (z / 30) - 1) < 0.8):
        triggered.add(71)
    if (abs((x / 50) * (y / 50) * (z / 30) - 1) < 0.3) != (abs((x / 80) * (y / 50) * (z / 30) - 1) < 0.3):
        triggered.add(72)

    # 73-78：均值联合与欧氏距离偏移
    if ((x + y) / 2 > 45 and (x + y) / 2 < 55 and z > 28 and z < 69) != (
            (x + y) / 2 > 55 and (x + y) / 2 < 55 and z > 38 and z < 69):
        triggered.add(73)
    if ((x + y) / 2 > 45 and (x + y) / 2 < 55 and z > 28 and z < 69) != (
            (x + y) / 2 > 55 and (x + y) / 2 < 53 and z > 48 and z < 69):
        triggered.add(74)
    if (abs(x - y) < 10 and z > 20 and z < 60) != (abs(x - y) < 20 and z > 20 and z < 60):
        triggered.add(75)
    if (abs(x - y) < 10 and z > 20 and z < 60) != (abs(x - y) < 10 and z > 40 and z < 60):
        triggered.add(76)
    if (math.sqrt((x - 50) ** 2 + (y - 50) ** 2) < 8 and abs(z - 30) < 3) != (
            math.sqrt((x - 50) ** 2 + (y - 50) ** 2) < 10 and abs(z - 30) < 3):
        triggered.add(77)
    if (math.sqrt((x - 50) ** 2 + (y - 50) ** 2) < 8 and abs(z - 30) < 3) != (
            math.sqrt((x - 50) ** 2 + (y - 50) ** 2) < 8 and abs(z - 30) < 10):
        triggered.add(78)

    # 79-87：归一化边界偏移
    if (x / 50 > 0.9 and x / 50 < 1.1 and y / 50 > 0.9 and y / 50 < 1.1) != (
            x / 50 > 1 and x / 50 < 1.1 and y / 50 > 0.9 and y / 50 < 1.1):
        triggered.add(79)
    if (x / 50 > 0.9 and x / 50 < 1.1 and y / 50 > 0.9 and y / 50 < 1.1) != (
            x / 50 > 1.2 and x / 50 < 1.1 and y / 50 > 0.9 and y / 50 < 1.1):
        triggered.add(80)
    if (z / 30 > 0.85 and z / 30 < 1) != (z / 30 > 0.85 and z / 30 < 1.3):
        triggered.add(81)
    if ((x / 50 + y / 50 + z / 30) / 3 > 0.9 and (x / 50 + y / 50 + z / 30) / 3 < 1) != (
            (x / 50 + y / 50 + z / 30) / 3 > 0.9 and (x / 50 + y / 50 + z / 30) / 3 < 1.5):
        triggered.add(82)
    if (max(abs(x / 50 - 1), abs(y / 50 - 1), abs(z / 30 - 1)) < 0.15) != (
            max(abs(x / 50 - 1), abs(y / 50 - 1), abs(z / 30 - 1)) < 0.25):
        triggered.add(83)
    if (max(abs(x / 50 - 1), abs(y / 50 - 1), abs(z / 30 - 1)) < 0.15) != (
            max(abs(x / 25 - 1), abs(y / 50 - 1), abs(z / 30 - 1)) < 0.15):
        triggered.add(84)
    if (min(x / 50, y / 50, z / 30) > 0.8) != (min(x / 50, y / 50, z / 30) > 1):
        triggered.add(85)
    if (min(x / 50, y / 50, z / 30) > 0.8) != (min(x / 50, y / 50, z / 30) > 1.8):
        triggered.add(86)
    if (max(x / 50, y / 50, z / 30) < 1.2) != (max(x / 50, y / 50, z / 30) < 1.6):
        triggered.add(87)

    # 88-94：缩放极差与复合公式偏移
    if (abs(max(x, y, z * 2) - min(x, y, z * 2)) < 20) != (abs(max(x, y, z * 3) - min(x, y, z * 3)) < 20):
        triggered.add(88)
    if (x + y + z * 1.5 > 100 and x + y + z * 1.5 < 300) != (x + y + z * 1.5 > 180 and x + y + z * 1.5 < 300):
        triggered.add(89)
    if (x * y * z > 40000 and x * y * z < 120000) != (x * y * z > 45000 and x * y * z < 120000):
        triggered.add(90)
    if (x * y * z > 40000 and x * y * z < 120000) != (x * y * z > 40000 and x * y * z < 95000):
        triggered.add(91)
    if (abs((x + y + z) / 3 - 45) < 3) != (abs((x + y + z) / 3 - 45) < 10):
        triggered.add(92)
    if (abs((x + y + z) / 3 - 45) < 3) != (abs((x + y + z) / 3 - 30) < 3):
        triggered.add(93)
    if (abs(x * y / z - 75) < 5) != (abs(x * y / z - 75) < 10):
        triggered.add(94)

    # 95-109：越界条件与极端值偏移
    if (x < 30 or y < 30 or z < 15) != (x < 20 or y < 30 or z < 15):
        triggered.add(95)
    if (x < 30 or y < 30 or z < 15) != (x < 30 or y < 20 or z < 15):
        triggered.add(96)
    if (x > 75 or y > 75 or z > 45) != (x > 60 or y > 75 or z > 45):
        triggered.add(97)
    if (x > 75 or y > 75 or z > 45) != (x > 75 or y > 60 or z > 45):
        triggered.add(98)
    if (abs(x / y - 1.0) > 0.25) != (abs(x / y - 1.0) > 0.3):
        triggered.add(99)
    if (abs(x / y - 1.0) > 0.25) != (abs(x / y - 0.95) > 0.25):
        triggered.add(100)
    if (abs(x / z - 1.8) > 0.3) != (abs(x / z - 1.8) > 0.4):
        triggered.add(101)
    if (abs(y / z - 1.7) > 0.3) != (abs(y / z - 1.6) > 0.3):
        triggered.add(102)
    if (x + y + z < 110 or x + y + z > 190) != (x + y + z < 115 or x + y + z > 190):
        triggered.add(103)
    if (x * y * z < 60000 or x * y * z > 140000) != (x * y * z < 60000 or x * y * z > 135000):
        triggered.add(104)
    if (x < 35 and y < 35) != (x < 30 and y < 35):
        triggered.add(105)
    if (x < 35 and y < 35) != (x < 35 and y < 30):
        triggered.add(106)
    if (y < 35 and z < 20) != (y < 30 and z < 20):
        triggered.add(107)
    if (x < 25 and y < 25 and z < 15) != (x < 22 and y < 25 and z < 15):
        triggered.add(108)
    if (x < 25 and y < 25 and z < 15) != (x < 25 and y < 22 and z < 15):
        triggered.add(109)

    return triggered


# === 绑定规则执行函数 ===
execute_Tr = section6_thermal_electrical_flow_hybrid

# ==================== 辅助函数 ====================
def clip_state(state):
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return np.clip(state, min_vals, max_vals)


def normalize_state(state):
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return 2 * (state - min_vals) / (max_vals - min_vals) - 1


def denormalize_state(normalized_state):
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return (normalized_state + 1) * (max_vals - min_vals) / 2 + min_vals


def coverage_similarity(triggered, target_path):
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0
    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)


def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']
    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']
    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']
    return reward


# ==================== SAC 网络结构 ====================
class GaussianPolicy(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dim=256):
        super(GaussianPolicy, self).__init__()
        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.mean = nn.Linear(hidden_dim, action_dim)
        self.log_std = nn.Linear(hidden_dim, action_dim)
        self.action_scale = 10.0  # 适配1~100的参数范围
        self.action_bias = 0.0

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))
        mean = self.mean(x)
        log_std = self.log_std(x)
        log_std = torch.clamp(log_std, min=-20, max=2)
        return mean, log_std

    def sample(self, state):
        mean, log_std = self.forward(state)
        std = log_std.exp()
        normal = Normal(mean, std)
        x_t = normal.rsample()
        action = torch.tanh(x_t) * self.action_scale + self.action_bias
        log_prob = normal.log_prob(x_t)
        log_prob -= torch.log(self.action_scale * (1 - torch.tanh(x_t).pow(2)) + 1e-6)
        log_prob = log_prob.sum(1, keepdim=True)
        mean = torch.tanh(mean) * self.action_scale + self.action_bias
        return action, log_prob, mean


class QNetwork(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dim=256):
        super(QNetwork, self).__init__()
        # Q1 网络
        self.fc1 = nn.Linear(state_dim + action_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, 1)
        # Q2 网络
        self.fc4 = nn.Linear(state_dim + action_dim, hidden_dim)
        self.fc5 = nn.Linear(hidden_dim, hidden_dim)
        self.fc6 = nn.Linear(hidden_dim, 1)

    def forward(self, state, action):
        xu = torch.cat([state, action], 1)
        x1 = F.relu(self.fc1(xu))
        x1 = F.relu(self.fc2(x1))
        q1 = self.fc3(x1)
        x2 = F.relu(self.fc4(xu))
        x2 = F.relu(self.fc5(x2))
        q2 = self.fc6(x2)
        return q1, q2


# ==================== 经验回放缓冲区 ====================
class EnhancedReplayBuffer:
    def __init__(self, capacity=50000):
        self.buffer = deque(maxlen=capacity)
        self.experience_info = deque(maxlen=capacity)

    def push(self, state, action, reward, next_state, done, path_idx, similarity):
        self.buffer.append((state, action, reward, next_state, done))
        self.experience_info.append({'path_idx': path_idx, 'similarity': similarity})

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return None
        batch = random.sample(self.buffer, batch_size)
        state, action, reward, next_state, done = map(np.stack, zip(*batch))
        return (
            torch.FloatTensor(state).to(device),
            torch.FloatTensor(action).to(device),
            torch.FloatTensor(reward).unsqueeze(1).to(device),
            torch.FloatTensor(next_state).to(device),
            torch.FloatTensor(done).unsqueeze(1).to(device)
        )

    def get_top_k_per_path(self, num_paths, k=20):
        path_samples = {i: [] for i in range(num_paths)}
        for idx, info in enumerate(self.experience_info):
            path_idx = info['path_idx']
            similarity = info['similarity']
            path_samples[path_idx].append((idx, similarity, self.buffer[idx]))

        top_k_results = {}
        for path_idx in range(num_paths):
            samples = path_samples[path_idx]
            if len(samples) == 0:
                top_k_results[path_idx] = []
                continue
            samples.sort(key=lambda x: x[1], reverse=True)
            top_k = samples[:k]
            top_k_results[path_idx] = []
            for sample in top_k:
                normalized_state = sample[2][0]
                original_state = denormalize_state(normalized_state)
                # 温度、电压取整，流量保留1位小数
                x, y, z = round(original_state[0]), round(original_state[1]), round(original_state[2], 1)
                triggered = execute_Tr(x, y, z)
                top_k_results[path_idx].append({
                    'state': np.array([x, y, z], dtype=np.float32),
                    'similarity': sample[1],
                    'triggered': triggered
                })
        return top_k_results

    def __len__(self):
        return len(self.buffer)


# ==================== SAC 智能体 ====================
class SACAgent:
    def __init__(self, state_dim=3, action_dim=3):
        self.state_dim = state_dim
        self.action_dim = action_dim
        # 策略网络
        self.policy = GaussianPolicy(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.policy_optimizer = optim.Adam(self.policy.parameters(), lr=EXPERIMENT_CONFIG['ACTOR_LR'])
        # 价值网络
        self.critic = QNetwork(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_target = QNetwork(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_target.load_state_dict(self.critic.state_dict())
        self.critic_optimizer = optim.Adam(self.critic.parameters(), lr=EXPERIMENT_CONFIG['CRITIC_LR'])
        # 自动熵调节
        self.target_entropy = -action_dim
        self.log_alpha = torch.zeros(1, requires_grad=True, device=device)
        self.alpha_optimizer = optim.Adam([self.log_alpha], lr=EXPERIMENT_CONFIG['ALPHA_LR'])
        # 经验池
        self.replay_buffer = EnhancedReplayBuffer()
        self.replay_train_count = 0

    def get_action(self, state, deterministic=False):
        normalized_state = normalize_state(state)
        state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)
        with torch.no_grad():
            if deterministic:
                _, _, action = self.policy.sample(state_tensor)
            else:
                action, _, _ = self.policy.sample(state_tensor)
        action = action.cpu().numpy()[0]
        return action

    def store_experience(self, state, action, reward, next_state, done, path_idx, similarity):
        normalized_state = normalize_state(state)
        normalized_next_state = normalize_state(next_state)
        self.replay_buffer.push(
            normalized_state, action, reward,
            normalized_next_state, done, path_idx, similarity
        )

    def replay_train(self):
        batch = self.replay_buffer.sample(EXPERIMENT_CONFIG['REPLAY_BATCH_SIZE'])
        if batch is None:
            return
        state, action, reward, next_state, done = batch

        with torch.no_grad():
            next_action, next_log_prob, _ = self.policy.sample(next_state)
            q1_next, q2_next = self.critic_target(next_state, next_action)
            q_next = torch.min(q1_next, q2_next)
            target_q = reward + (1 - done) * EXPERIMENT_CONFIG['GAMMA'] * (
                    q_next - self.log_alpha.exp() * next_log_prob)

        # 更新 Critic
        q1, q2 = self.critic(state, action)
        critic_loss = F.mse_loss(q1, target_q) + F.mse_loss(q2, target_q)
        self.critic_optimizer.zero_grad()
        critic_loss.backward()
        self.critic_optimizer.step()

        # 更新 Actor
        new_action, log_prob, _ = self.policy.sample(state)
        q1_new, q2_new = self.critic(state, new_action)
        q_new = torch.min(q1_new, q2_new)
        policy_loss = (self.log_alpha.exp() * log_prob - q_new).mean()
        self.policy_optimizer.zero_grad()
        policy_loss.backward()
        self.policy_optimizer.step()

        # 更新 Alpha
        alpha_loss = -(self.log_alpha * (log_prob + self.target_entropy).detach()).mean()
        self.alpha_optimizer.zero_grad()
        alpha_loss.backward()
        self.alpha_optimizer.step()

        # 软更新目标网络
        for param, target_param in zip(self.critic.parameters(), self.critic_target.parameters()):
            target_param.data.copy_(
                EXPERIMENT_CONFIG['TAU'] * param.data + (1 - EXPERIMENT_CONFIG['TAU']) * target_param.data)

        self.replay_train_count += 1
        if self.replay_train_count % 2 == 0:
            alpha_value = self.log_alpha.exp().item()
            print(f"  -> SAC 更新 (第 {self.replay_train_count} 次), Alpha={alpha_value:.4f}")


# ==================== 性能指标计算 ====================
def calculate_run_performance(run_idx, sac_results, training_time, total_steps, update_count, agent):
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)
    all_similarities = []
    total_samples = 0
    total_reward = 0

    for path_idx in range(num_paths):
        samples = sac_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']
            total_reward += reward
            all_similarities.append(similarity)
            total_samples += 1

    average_reward = total_reward / total_samples if total_samples > 0 else 0
    convergence = np.mean(all_similarities) if all_similarities else 0
    environment_adaptability = 1 / (np.std(all_similarities) + 1e-8) if len(all_similarities) > 1 else 0
    generalization_ability = convergence
    computational_efficiency = total_steps / training_time if training_time > 0 else 0
    policy_update_frequency = update_count / training_time if training_time > 0 else 0
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0

    return {
        'Run': run_idx + 1,
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }


# ==================== Excel 导出 ====================
def export_to_excel(all_sac_results, all_performance_data, target_paths, output_path="SAC_20_run.xlsx"):
    print("\n正在导出数据到 Excel...")
    all_sac_summary_data = []
    all_sac_detailed_data = []

    for run_idx, (sac_results, performance_data) in enumerate(zip(all_sac_results, all_performance_data)):
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = sac_results[path_idx]
            if len(samples) == 0:
                all_sac_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue
            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = 'Yes' if perfect_count > 0 else 'No'
            all_sac_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })

            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']
                all_sac_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Sample ID': sample_idx + 1,
                    'X(温度)': int(state[0]),
                    'Y(电压)': int(state[1]),
                    'Z(流量)': float(state[2]),
                    'Similarity': round(similarity, 4),
                    'Perfect Coverage': 'Yes' if abs(similarity - 1.0) < 0.001 else 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path))),
                    'Triggered Rules': ', '.join(map(str, sorted(triggered))),
                    'Hit Rule Count': len(target_path.intersection(triggered)),
                    'Target Rule Count': len(target_path)
                })

    summary_df = pd.DataFrame(all_sac_summary_data)
    detailed_df = pd.DataFrame(all_sac_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='SACPath', index=False)
        detailed_df.to_excel(writer, sheet_name='SACDetailed Sample Data', index=False)
        selected_columns = [
            'Run', 'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df[selected_columns].to_excel(writer, sheet_name='Metric', index=False)

        # 格式美化
        workbook = writer.book
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        perfect_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        for sheet_name in ['SACPath', 'SACDetailed Sample Data', 'Metric']:
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 完美覆盖行标绿
        ws1 = writer.sheets['SACPath']
        for row_idx in range(2, ws1.max_row + 1):
            if ws1.cell(row_idx, 9).value == 'Yes':
                for col_idx in range(1, ws1.max_column + 1):
                    ws1.cell(row_idx, col_idx).fill = perfect_fill

    print(f"文件已保存到: {output_path}")
    print(f"  - Sheet1: SACPath 统计 ({len(all_sac_summary_data)} 条)")
    print(f"  - Sheet2: SACDetailed Sample Data ({len(all_sac_detailed_data)} 条)")
    print(f"  - Sheet3: Metric ({len(all_performance_data)} 条)")


# ==================== 单次训练流程 ====================
def train_sac_workflow():
    print("=" * 80)
    print("SAC 训练 - 热电流量混合控制 (section6)")
    print("状态范围: X温度 1~100, Y电压 1~100, Z流量 1~60")
    print("相似度 = 命中规则数 / 目标路径规则数")
    print("=" * 80)

    agent = SACAgent()
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n生成初始样本: {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']} / 路径")
    path_samples = {}
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']

    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            state = np.random.uniform(min_vals, max_vals).astype(np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  路径 {path_idx + 1}/{num_paths}: {len(samples)} 个初始样本")

    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size

    print(f"\n批次大小: {batch_size}, 单样本迭代步数: {EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(f"总批次数: {num_batches} /路径 × {num_paths} 路径 = {num_batches * num_paths}")
    print("-" * 80)

    for batch_idx in range(num_batches):
        print(f"\n批次 {batch_idx + 1}/{num_batches}")

        for path_idx in range(num_paths):
            target_path = target_paths[path_idx]
            batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]

            batch_rewards = []
            batch_similarities = []

            for initial_state in batch_samples:
                state = initial_state.copy()
                episode_reward = 0
                final_similarity = 0

                for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                    action = agent.get_action(state, deterministic=False)
                    next_state = state + action
                    next_state = clip_state(next_state)

                    x, y, z = next_state
                    triggered = execute_Tr(x, y, z)
                    reward = unified_reward_function(triggered, target_path)
                    similarity = coverage_similarity(triggered, target_path)

                    done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)

                    agent.store_experience(
                        state, action, reward, next_state, done,
                        path_idx, similarity
                    )

                    state = next_state
                    episode_reward += reward
                    final_similarity = similarity
                    total_steps += 1

                batch_rewards.append(episode_reward)
                batch_similarities.append(final_similarity)

            avg_reward = np.mean(batch_rewards)
            avg_similarity = np.mean(batch_similarities)
            print(f"  路径 {path_idx + 1}: 平均奖励={avg_reward:.2f}, 平均相似度={avg_similarity:.4f}")

        print(f"\n  执行 SAC 更新...")
        agent.replay_train()
        print(f"  经验池大小: {len(agent.replay_buffer)}")

    training_time = time.time() - start_time

    print("\n" + "=" * 80)
    print(f"单次运行训练完成! 总耗时: {training_time:.2f} 秒, 总交互步数: {total_steps}")
    print(f"经验池大小: {len(agent.replay_buffer)}")
    print(f"SAC 更新次数: {agent.replay_train_count}")
    print("=" * 80)

    print(f"\n提取每条路径的 Top-{EXPERIMENT_CONFIG['TOP_K_SAMPLES']} 高相似度样本...")
    top_k_results = agent.replay_buffer.get_top_k_per_path(num_paths, EXPERIMENT_CONFIG['TOP_K_SAMPLES'])

    return agent, top_k_results, training_time, total_steps, agent.replay_train_count


# ==================== 主程序 ====================
def main():
    print("\n" + "=" * 80)
    print("SAC 多轮实验 - 热电流量混合控制")
    print("参数范围: X温度 1~100, Y电压 1~100, Z流量 1~60")
    print("=" * 80)

    all_sac_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'='*80}")
        print(f"开始第 {run_idx + 1}/{EXPERIMENT_CONFIG['NUM_RUNS']} 次独立运行")
        print(f"{'='*80}")

        sac_agent, sac_results, training_time, total_steps, update_count = train_sac_workflow()
        performance_data = calculate_run_performance(
            run_idx, sac_results, training_time, total_steps, update_count, sac_agent
        )

        all_sac_results.append(sac_results)
        all_performance_data.append(performance_data)

        print(f"\n第 {run_idx + 1} 次运行完成!")
        print(f"  总奖励: {performance_data['Total Reward']}")
        print(f"  平均奖励: {performance_data['Average Reward']}")
        print(f"  平均相似度: {performance_data['Average Similarity']}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"SAC_thermal_electrical_flow_20run_{timestamp}.xlsx"
    export_to_excel(all_sac_results, all_performance_data, target_paths, output_path)

    # 整体统计
    print("\n" + "=" * 80)
    print(f"{EXPERIMENT_CONFIG['NUM_RUNS']} 次运行整体统计摘要")
    print("=" * 80)

    metrics = ['Total Reward', 'Average Reward', 'Convergence',
               'Environment Adaptability', 'Generalization Ability',
               'Computational Efficiency', 'Policy Update Frequency', 'Average Similarity']

    for metric in metrics:
        values = [p[metric] for p in all_performance_data]
        print(f"\n{metric} Statistics:")
        print(f"  Mean: {np.mean(values):.4f}")
        print(f"  Std: {np.std(values):.4f}")

    print("\n" + "=" * 80)
    print(f" {EXPERIMENT_CONFIG['NUM_RUNS']} 次运行全部完成!")
    print("=" * 80)


if __name__ == "__main__":
    main()