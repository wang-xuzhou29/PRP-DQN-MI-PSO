import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

def execute_validation_rules(x, y, z):
    # 确保数值在范围内
    x = max(STATE_MIN_X, min(STATE_MAX_X, x))
    y = max(STATE_MIN_Y, min(STATE_MAX_Y, y))
    z = max(STATE_MIN_Z, min(STATE_MAX_Z, z))

    triggered = set()

    if (25 < x < 75) != (25 < x < 80):
        triggered.add(1)
    if (25 < x < 75) != (30 < x < 75):
        triggered.add(2)

    if (25 < y < 75) != (40 < y < 75):
        triggered.add(3)
    if (25 < y < 75) != (50 < y < 75):
        triggered.add(4)

    if (20 < z < 50) != (30 < z < 50):
        triggered.add(5)
    if (20 < z < 50) != (20 < z < 30):
        triggered.add(6)

    if (x > 40 and x < 60 and y > 35 and y < 65) != (x > 40 and x < 75 and y > 35 and y < 65):
        triggered.add(7)
    if (x > 40 and x < 60 and y > 35 and y < 65) != (x > 40 and x < 60 and y > 20 and y < 65):
        triggered.add(8)

    if (x > 45 and x < 55 and z > 25 and z < 35) != (x > 45 and x < 55 and z > 10 and z < 35):
        triggered.add(9)

    if (y > 40 and y < 60 and z > 25 and z < 35) != (y > 40 and y < 85 and z > 25 and z < 35):
        triggered.add(10)
    if (y > 40 and y < 60 and z > 25 and z < 35) != (y > 40 and y < 60 and z > 30 and z < 40):
        triggered.add(11)

    if (x > 30 and x < 52) != (x > 30 and x < 70):
        triggered.add(12)
    if (x > 30 and x < 52) != (x > 30 and x < 80):
        triggered.add(13)

    if (y > 30 and y < 52) != (y > 30 and y < 70):
        triggered.add(14)
    if (y > 30 and y < 52) != (y > 30 and y < 80):
        triggered.add(15)

    if (z > 20 and z < 32) != (z > 20 and z < 40):
        triggered.add(16)
    if (z > 20 and z < 32) != (z > 20 and z < 50):
        triggered.add(17)

    if (abs(x - 50) < 5) != (abs(x - 50) < 10):
        triggered.add(18)
    if (abs(x - 50) < 5) != (abs(x - 30) < 5):
        triggered.add(19)

    if (abs(y - 50) < 8) != (abs(y - 50) < 15):
        triggered.add(20)
    if (abs(y - 50) < 8) != (abs(y - 20) < 8):
        triggered.add(21)

    if (abs(z - 30) < 3) != (abs(z - 30) < 6):
        triggered.add(22)
    if (abs(z - 30) < 3) != (abs(z - 20) < 3):
        triggered.add(23)

    if (x > 35 and y > 35 and z > 20) != (x > 35 and y > 60 and z > 20):
        triggered.add(24)

    if (x < 75 and y < 75 and z < 45) != (x < 75 and y < 40 and z < 45):
        triggered.add(25)
    if (x < 75 and y < 75 and z < 45) != (x < 75 and y < 30 and z < 45):
        triggered.add(26)

    if (x + y + z > 20 and x + y + z < 180) != (x + y + z > 50 and x + y + z < 180):
        triggered.add(27)
    if (x + y + z > 20 and x + y + z < 180) != (x + y + z > 60 and x + y + z < 180):
        triggered.add(28)

    if (x * y > 1000 and x * y < 4000) != (x * y > 1500 and x * y < 4000):
        triggered.add(29)
    if (x * y > 1000 and x * y < 4000) != (x * y > 1000 and x * y < 5000):
        triggered.add(30)

    if (x * z > 1200 and x * z < 3400) != (x * z > 1500 and x * z < 3400):
        triggered.add(31)
    if (x * z > 1200 and x * z < 3400) != (x * z > 1200 and x * z < 3000):
        triggered.add(32)

    if (y * z > 1000 and y * z < 3000) != (y * z > 1500 and y * z < 3000):
        triggered.add(33)
    if (y * z > 1000 and y * z < 3000) != (y * z > 1000 and y * z < 3500):
        triggered.add(34)

    if ((x + y + z) / 3 > 35 and (x + y + z) / 3 < 55) != ((x + y + z) / 3 > 40 and (x + y + z) / 3 < 55):
        triggered.add(35)
    if ((x + y + z) / 3 > 35 and (x + y + z) / 3 < 55) != ((x + y + z) / 3 > 35 and (x + y + z) / 3 < 40):
        triggered.add(36)

    if (math.sqrt(x ** 2 + y ** 2 + z ** 2) > 180) != (math.sqrt(x ** 2 + y ** 2 + z ** 2) > 100):
        triggered.add(37)
    if (math.sqrt(x ** 2 + y ** 2 + z ** 2) > 180) != (math.sqrt(x ** 2 + y ** 2 + z ** 2) > 120):
        triggered.add(38)

    if (x / y > 1 and x / y < 3) != (x / y > 1 and x / y < 1.2):
        triggered.add(39)
    if (x / y > 1 and x / y < 3) != (x / y > 1 and x / y < 2):
        triggered.add(40)

    if (x / z > 1.5 and x / z < 4.5) != (x / z > 1.8 and x / z < 4.5):
        triggered.add(41)
    if (x / z > 1.5 and x / z < 4.5) != (x / z > 1.5 and x / z < 3):
        triggered.add(42)

    if (y / z > 1.2 and y / z < 4.0) != (y / z > 1.4 and y / z < 4.0):
        triggered.add(43)
    if (y / z > 1.2 and y / z < 4.0) != (y / z > 1.2 and y / z < 3.5):
        triggered.add(44)

    if ((x - 50) + (y - 50) > 10 and (x - 50) - (y - 50) < 10) != (
            (x - 50) + (y - 50) > 20 and (x - 50) - (y - 50) < 10):
        triggered.add(45)
    if ((x - 50) + (y - 50) > 10 and (x - 50) - (y - 50) < 10) != (
            (x - 50) + (y - 50) > 50 and (x - 50) - (y - 50) < 10):
        triggered.add(46)

    if ((x - 50) + (z - 30) * 2 > 15 and (x - 50) + (z - 30) * 2 < 15) != (
            (x - 50) + (z - 30) * 2 > 10 and (x - 50) + (z - 30) * 2 < 15):
        triggered.add(47)
    if ((x - 50) + (z - 30) * 2 > 15 and (x - 50) + (z - 30) * 2 < 15) != (
            (x - 50) + (z - 30) * 2 > 15 and (x - 50) + (z - 30) * 2 < 40):
        triggered.add(48)

    if ((y - 50) + (z - 30) > 20 and (y - 50) + (z - 30) < 20) != (
            (y - 50) + (z - 30) > 10 and (y - 50) + (z - 30) < 20):
        triggered.add(49)
    if ((y - 50) + (z - 30) > 20 and (y - 50) + (z - 30) < 20) != (
            (y - 50) + (z - 30) > 20 and (y - 50) + (z - 30) < 40):
        triggered.add(50)

    if (abs((x - 50) - (y - 50)) < 8) != (abs((x - 50) - (y - 50)) < 30):
        triggered.add(51)
    if (abs((x - 50) - (y - 50)) < 8) != (abs((x - 45) - (y - 50)) < 20):
        triggered.add(52)

    if (abs((x - 50) - (z - 30) * 2) < 10) != (abs((x - 50) - (z - 30) * 2) < 30):
        triggered.add(53)
    if (abs((x - 50) - (z - 30) * 2) < 10) != (abs((x - 80) - (z - 30) * 2) < 10):
        triggered.add(54)

    if (abs((y - 50) - (z - 30)) < 5) != (abs((y - 50) - (z - 30)) < 20):
        triggered.add(55)
    if (abs((y - 50) - (z - 30)) < 5) != (abs((y - 80) - (z - 30)) < 5):
        triggered.add(56)

    if (x / (y + 10) > 0.7 and x / (y + 10) < 1.1) != (x / (y + 10) > 0.8 and x / (y + 10) < 1.5):
        triggered.add(57)
    if (x / (y + 10) > 0.7 and x / (y + 10) < 1.1) != (x / (y + 10) > 0.5 and x / (y + 10) < 1.0):
        triggered.add(58)

    if (y / (z + 20) > 1.0 and y / (z + 20) < 1.8) != (y / (z + 20) > 1.5 and y / (z + 20) < 1.8):
        triggered.add(59)

    if (z / (x / 2) > 0.8 and z / (x / 2) < 1.4) != (z / (x / 2) > 1.5 and z / (x / 2) < 1.4):
        triggered.add(60)
    if (z / (x / 2) > 0.8 and z / (x / 2) < 1.4) != (z / (x / 2) > 0.8 and z / (x / 2) < 1.9):
        triggered.add(61)

    if (x * y * z > 50000 and x * y * z < 150000) != (x * y * z > 30000 and x * y * z < 150000):
        triggered.add(62)

    if (x * 0.4 + y * 0.4 + z * 0.8 > 35 and x * 0.4 + y * 0.4 + z * 0.8 < 65) != (
            x * 0.4 + y * 0.4 + z * 0.8 > 35 and x * 0.4 + y * 0.4 + z * 0.8 < 60):
        triggered.add(63)

    if ((x * y * z) ** (1 / 3) > 25) != ((x * y * z) ** (1 / 4) > 25):
        triggered.add(64)
    if ((x * y * z) ** (1 / 3) > 25) != ((x * y * z) ** (1 / 2) > 25):
        triggered.add(65)

    if ((x - 50) * (y - 50) > 200 and (x - 50) * (y - 50) < 200) != (
            (x - 50) * (y - 50) > 150 and (x - 50) * (y - 50) < 200):
        triggered.add(66)

    if ((x - 50) * (z - 30) > 150 and (x - 50) * (z - 30) < 150) != (
            (x - 50) * (z - 30) > 100 and (x - 50) * (z - 30) < 150):
        triggered.add(67)
    if ((x - 50) * (z - 30) > 150 and (x - 50) * (z - 30) < 150) != (
            (x - 30) * (z - 30) > 150 and (x - 50) * (z - 30) < 150):
        triggered.add(68)

    if (abs(x / 50 + y / 50 + z / 30 - 4) < 0.5) != (abs(x / 50 + y / 50 + z / 30 - 4) < 1):
        triggered.add(69)
    if (abs(x / 50 + y / 50 + z / 30 - 4) < 0.5) != (abs(x / 40 + y / 50 + z / 30 - 4) < 0.5):
        triggered.add(70)

    if (abs((x / 50) * (y / 50) * (z / 30) - 1) < 0.3) != (abs((x / 50) * (y / 50) * (z / 30) - 1) < 0.8):
        triggered.add(71)
    if (abs((x / 50) * (y / 50) * (z / 30) - 1) < 0.3) != (abs((x / 80) * (y / 50) * (z / 30) - 1) < 0.3):
        triggered.add(72)

    if ((x + y) / 2 > 45 and (x + y) / 2 < 55 and z > 28 and z < 69) != (
            (x + y) / 2 > 55 and (x + y) / 2 < 55 and z > 38 and z < 69):
        triggered.add(73)
    if ((x + y) / 2 > 45 and (x + y) / 2 < 55 and z > 28 and z < 69) != (
            (x + y) / 2 > 55 and (x + y) / 2 < 53 and z > 48 and z < 69):
        triggered.add(74)

    if (abs(x - y) < 10 and z > 20 and z < 60) != (abs(x - y) < 20 and z > 20 and z < 60):
        triggered.add(75)
    if (abs(x - y) < 10 and z > 20 and z < 60) != (abs(x - y) < 10 and z > 40 and z < 60):
        triggered.add(76)

    if (math.sqrt((x - 50) ** 2 + (y - 50) ** 2) < 8 and abs(z - 30) < 3) != (
            math.sqrt((x - 50) ** 2 + (y - 50) ** 2) < 10 and abs(z - 30) < 3):
        triggered.add(77)
    if (math.sqrt((x - 50) ** 2 + (y - 50) ** 2) < 8 and abs(z - 30) < 3) != (
            math.sqrt((x - 50) ** 2 + (y - 50) ** 2) < 8 and abs(z - 30) < 10):
        triggered.add(78)

    if (x / 50 > 0.9 and x / 50 < 1.1 and y / 50 > 0.9 and y / 50 < 1.1) != (
            x / 50 > 1 and x / 50 < 1.1 and y / 50 > 0.9 and y / 50 < 1.1):
        triggered.add(79)
    if (x / 50 > 0.9 and x / 50 < 1.1 and y / 50 > 0.9 and y / 50 < 1.1) != (
            x / 50 > 1.2 and x / 50 < 1.1 and y / 50 > 0.9 and y / 50 < 1.1):
        triggered.add(80)

    if (z / 30 > 0.85 and z / 30 < 1) != (z / 30 > 0.85 and z / 30 < 1.3):
        triggered.add(81)

    if ((x / 50 + y / 50 + z / 30) / 3 > 0.9 and (x / 50 + y / 50 + z / 30) / 3 < 1) != (
            (x / 50 + y / 50 + z / 30) / 3 > 0.9 and (x / 50 + y / 50 + z / 30) / 3 < 1.5):
        triggered.add(82)

    if (max(abs(x / 50 - 1), abs(y / 50 - 1), abs(z / 30 - 1)) < 0.15) != (
            max(abs(x / 50 - 1), abs(y / 50 - 1), abs(z / 30 - 1)) < 0.25):
        triggered.add(83)
    if (max(abs(x / 50 - 1), abs(y / 50 - 1), abs(z / 30 - 1)) < 0.15) != (
            max(abs(x / 25 - 1), abs(y / 50 - 1), abs(z / 30 - 1)) < 0.15):
        triggered.add(84)

    if (min(x / 50, y / 50, z / 30) > 0.8) != (min(x / 50, y / 50, z / 30) > 1):
        triggered.add(85)
    if (min(x / 50, y / 50, z / 30) > 0.8) != (min(x / 50, y / 50, z / 30) > 1.8):
        triggered.add(86)

    if (max(x / 50, y / 50, z / 30) < 1.2) != (max(x / 50, y / 50, z / 30) < 1.6):
        triggered.add(87)

    if (abs(max(x, y, z * 2) - min(x, y, z * 2)) < 20) != (abs(max(x, y, z * 3) - min(x, y, z * 3)) < 20):
        triggered.add(88)

    if (x + y + z * 1.5 > 100 and x + y + z * 1.5 < 300) != (x + y + z * 1.5 > 180 and x + y + z * 1.5 < 300):
        triggered.add(89)

    if (x * y * z > 40000 and x * y * z < 120000) != (x * y * z > 45000 and x * y * z < 120000):
        triggered.add(90)
    if (x * y * z > 40000 and x * y * z < 120000) != (x * y * z > 40000 and x * y * z < 95000):
        triggered.add(91)

    if (abs((x + y + z) / 3 - 45) < 3) != (abs((x + y + z) / 3 - 45) < 10):
        triggered.add(92)
    if (abs((x + y + z) / 3 - 45) < 3) != (abs((x + y + z) / 3 - 30) < 3):
        triggered.add(93)

    if (abs(x * y / z - 75) < 5) != (abs(x * y / z - 75) < 10):
        triggered.add(94)

    if (x < 30 or y < 30 or z < 15) != (x < 20 or y < 30 or z < 15):
        triggered.add(95)
    if (x < 30 or y < 30 or z < 15) != (x < 30 or y < 20 or z < 15):
        triggered.add(96)

    if (x > 75 or y > 75 or z > 45) != (x > 60 or y > 75 or z > 45):
        triggered.add(97)
    if (x > 75 or y > 75 or z > 45) != (x > 75 or y > 60 or z > 45):
        triggered.add(98)

    if (abs(x / y - 1.0) > 0.25) != (abs(x / y - 1.0) > 0.3):
        triggered.add(99)
    if (abs(x / y - 1.0) > 0.25) != (abs(x / y - 0.95) > 0.25):
        triggered.add(100)

    if (abs(x / z - 1.8) > 0.3) != (abs(x / z - 1.8) > 0.4):
        triggered.add(101)

    if (abs(y / z - 1.7) > 0.3) != (abs(y / z - 1.6) > 0.3):
        triggered.add(102)

    if (x + y + z < 110 or x + y + z > 190) != (x + y + z < 115 or x + y + z > 190):
        triggered.add(103)

    if (x * y * z < 60000 or x * y * z > 140000) != (x * y * z < 60000 or x * y * z > 135000):
        triggered.add(104)

    if (x < 35 and y < 35) != (x < 30 and y < 35):
        triggered.add(105)
    if (x < 35 and y < 35) != (x < 35 and y < 30):
        triggered.add(106)

    if (y < 35 and z < 20) != (y < 30 and z < 20):
        triggered.add(107)

    if (x < 25 and y < 25 and z < 15) != (x < 22 and y < 25 and z < 15):
        triggered.add(108)
    if (x < 25 and y < 25 and z < 15) != (x < 25 and y < 22 and z < 15):
        triggered.add(109)

    return triggered

def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """"""
    generated_path = execute_validation_rules(particle[0], particle[1], particle[2])

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # : x:1-50, y:1-50, z:1-50
        self.bounds = bounds if bounds else [(1, 100), (1, 100), (1, 60)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """target pathsPSO"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(particles[i][0], particles[i][1], particles[i][2]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(gbest_particle[0], gbest_particle[1], gbest_particle[2]),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """Path PSO"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - Path ")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations")
    print(f"Path : {len(target_paths)}")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"Path {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "" if result['success'] else f"({result['best_fitness']:.3f})"
        print(f"{status} | {result['time']:.2f}s | iterations{result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f": {success_count}/{len(target_paths)} ({success_rate:.1f}%) | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs}")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations, {len(target_paths)}Path ")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"---  {run_idx}/{num_runs} ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f": {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs} runcompleted | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 1: 
    ws1 = wb.active
    ws1.title = ""
    ws1.sheet_view.showGridLines = False

    headers = ["Run", "", "", "", "Average Iterations", "(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f" {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # 2: Path 
    ws2 = wb.create_sheet(title="Path ")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Path ID", "", "", "", "Average Iterations", "Minimum Iterations", "Maximum Iterations"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"Path  {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # 3: 
    ws3 = wb.create_sheet(title="")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Path ", "", "(x,y,z)", "", "Iterations", "Path "]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"{run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # 4: target paths
    ws4 = wb.create_sheet(title="target paths")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Path ID", "target paths", ""]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"Path  {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f" : {filename}")
    print(f"{'=' * 70}")
    print(f":")
    print(f"  1.        - {len(all_results)} run")
    print(f"  2. Path        - Path ")
    print(f"  3.    -  runPath ")
    print(f"  4. target paths       - target paths")
    print(f"{'=' * 70}\n")

    return filename


def main():


    target_paths = [
        {4, 6, 11, 12, 13, 16, 17, 18, 21, 24, 25, 26, 36, 39, 41, 43, 48, 51, 52, 54, 55, 58, 60, 64, 68, 69, 70, 71,
         72,
         73, 74, 75, 81, 82, 83, 85, 86, 87, 89, 91, 93, 99, 102},
        {6, 10, 12, 13, 14, 15, 16, 17, 18, 20, 22, 25, 26, 36, 41, 45, 46, 47, 49, 51, 52, 54, 55, 57, 59, 60, 64, 69,
         70,
         71, 72, 76, 81, 82, 86, 87, 88, 89, 91, 92, 98, 102},
        {6, 7, 11, 12, 13, 14, 15, 16, 17, 21, 24, 25, 26, 36, 39, 41, 43, 48, 49, 51, 52, 54, 55, 58, 60, 64, 67, 68,
         72,
         75, 82, 86, 87, 89, 92, 97, 100, 102, 104},
        {6, 11, 12, 13, 14, 15, 16, 17, 19, 21, 24, 25, 26, 36, 48, 49, 53, 56, 57, 61, 64, 68, 69, 70, 71, 72, 73, 74,
         76,
         78, 81, 82, 83, 86, 87, 89, 91, 92, 101},
        {6, 7, 13, 14, 15, 16, 17, 20, 22, 24, 25, 26, 30, 36, 39, 48, 49, 51, 52, 53, 56, 58, 59, 60, 64, 66, 68, 72,
         75,
         81, 82, 86, 87, 88, 89, 92, 97, 100, 104},
        {4, 6, 12, 13, 16, 17, 19, 21, 22, 24, 25, 26, 33, 36, 39, 41, 43, 47, 51, 52, 54, 55, 60, 64, 69, 72, 73, 74,
         76,
         78, 81, 82, 84, 85, 86, 89, 93, 94, 100},
        {5, 7, 11, 12, 13, 14, 15, 21, 22, 24, 25, 26, 33, 36, 39, 47, 51, 52, 53, 54, 55, 58, 59, 64, 69, 70, 71, 72,
         75,
         82, 85, 86, 87, 88, 89, 91, 92, 97, 100},
        {6, 10, 14, 15, 16, 17, 18, 20, 22, 24, 25, 26, 31, 36, 49, 51, 52, 53, 55, 58, 59, 61, 64, 69, 73, 74, 75, 81,
         82,
         83, 85, 86, 87, 89, 93, 94, 99, 100},
        {5, 7, 11, 13, 14, 15, 21, 23, 24, 25, 26, 30, 36, 39, 48, 51, 53, 54, 55, 58, 59, 60, 64, 66, 71, 72, 75, 82,
         85,
         86, 87, 88, 89, 91, 92, 97, 99, 102},
        {6, 10, 12, 13, 15, 16, 17, 19, 23, 25, 26, 36, 41, 46, 50, 51, 52, 54, 55, 58, 59, 60, 64, 71, 72, 75, 81, 82,
         86,
         87, 88, 89, 91, 92, 98, 99, 100},
        {4, 6, 11, 12, 13, 16, 17, 19, 21, 22, 24, 25, 26, 36, 43, 47, 54, 55, 60, 64, 69, 72, 73, 74, 76, 78, 80, 81,
         82,
         83, 85, 86, 89, 93, 94, 101, 102},
        {5, 6, 11, 14, 15, 19, 20, 23, 24, 25, 26, 31, 36, 41, 51, 52, 54, 55, 57, 59, 60, 64, 69, 72, 73, 74, 75, 77,
         81,
         82, 83, 85, 86, 88, 89, 93, 102},
        {5, 6, 11, 14, 15, 19, 21, 23, 24, 25, 26, 31, 36, 41, 54, 56, 57, 59, 60, 63, 64, 69, 72, 73, 74, 76, 79, 80,
         81,
         82, 84, 85, 86, 88, 89, 93, 94},
        {5, 6, 7, 10, 13, 14, 15, 20, 23, 25, 26, 30, 37, 46, 48, 49, 51, 52, 53, 54, 55, 59, 60, 64, 72, 76, 81, 82,
         85,
         86, 87, 88, 89, 97, 98, 104},
        {5, 9, 12, 13, 14, 15, 19, 20, 25, 26, 31, 33, 36, 45, 46, 51, 52, 53, 55, 57, 59, 60, 63, 64, 69, 72, 75, 82,
         83,
         87, 88, 89, 93, 98, 101},
        {3, 4, 6, 7, 12, 13, 16, 17, 20, 22, 24, 26, 33, 36, 39, 48, 51, 54, 55, 57, 60, 64, 68, 69, 70, 72, 73, 74, 81,
         82,
         87, 89, 93, 94, 97},
        {1, 5, 10, 13, 14, 15, 20, 22, 30, 36, 37, 39, 48, 51, 52, 54, 55, 58, 59, 64, 71, 72, 75, 82, 85, 86, 87, 88,
         89,
         92, 97, 98, 100},
        {5, 7, 13, 14, 15, 20, 25, 26, 30, 33, 36, 42, 47, 51, 52, 54, 55, 58, 59, 64, 69, 70, 71, 72, 75, 82, 87, 88,
         89,
         91, 92, 97, 98},
        {3, 4, 6, 16, 17, 19, 20, 23, 24, 26, 33, 35, 36, 39, 41, 43, 51, 52, 54, 55, 58, 60, 63, 64, 72, 75, 81, 83,
         89,
         92, 99},
        {12, 13, 15, 23, 25, 26, 30, 31, 33, 36, 37, 42, 44, 46, 49, 54, 56, 64, 69, 70, 71, 72, 82, 87, 88, 89, 91, 92,
         97,
         98},
        {6, 7, 13, 14, 15, 17, 20, 24, 30, 32, 37, 39, 41, 43, 50, 51, 52, 54, 55, 58, 60, 64, 70, 75, 82, 86, 87, 97,
         100,
         102},
        {6, 12, 13, 15, 17, 18, 30, 34, 37, 46, 48, 50, 51, 52, 53, 55, 58, 59, 61, 64, 66, 68, 82, 86, 87, 99, 100,
         101,
         102},
        {3, 4, 6, 13, 16, 17, 22, 26, 33, 36, 39, 40, 48, 53, 60, 64, 68, 69, 70, 72, 73, 74, 81, 82, 87, 89, 93, 97,
         101},
        {3, 4, 6, 8, 12, 13, 16, 17, 19, 21, 22, 29, 35, 39, 40, 47, 51, 54, 57, 60, 62, 64, 71, 81, 89, 92, 96, 101,
         103},
        {2, 6, 10, 14, 15, 16, 17, 19, 22, 25, 26, 36, 50, 53, 55, 59, 63, 64, 72, 73, 74, 81, 82, 87, 89, 93, 95, 98},
        {3, 4, 5, 9, 12, 13, 19, 20, 23, 24, 26, 35, 39, 51, 52, 53, 56, 58, 60, 62, 64, 71, 75, 88, 89, 90, 92, 103},
        {6, 15, 16, 17, 37, 38, 39, 50, 51, 53, 54, 55, 58, 59, 60, 64, 69, 75, 81, 86, 88, 100, 102},
        {3, 4, 31, 35, 48, 49, 51, 52, 58, 62, 64, 69, 71, 82, 89, 90, 92, 95, 99, 100, 103, 106},
        {3, 4, 12, 13, 23, 26, 35, 39, 42, 51, 54, 55, 57, 62, 64, 71, 87, 89, 92, 103, 107},
        {3, 4, 6, 17, 19, 26, 29, 31, 33, 35, 47, 62, 63, 64, 71, 87, 89, 92, 103, 105, 106},
        {7, 12, 13, 14, 15, 21, 25, 26, 35, 39, 51, 52, 56, 58, 65, 87, 89, 92, 97, 100},
        {3, 4, 26, 28, 51, 52, 53, 54, 55, 58, 59, 65, 99, 100, 106, 107},
        {3, 4, 26, 27, 28, 41, 51, 52, 54, 55, 59, 60, 65, 106, 107},
        {27, 28, 39, 40, 42, 51, 52, 53, 54, 55, 57, 65, 108},
        {21, 27, 28, 53, 54, 56, 58, 59, 65, 88, 109},
    ]

    print("=" * 70)
    print("baseline PSO")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("Program completed")


if __name__ == "__main__":
    main()