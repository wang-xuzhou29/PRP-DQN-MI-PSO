import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

NUM_RUNS = 20  # Experiment Runs, 

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === new three-dimensional range settings ===
LIGHT_MIN = 1
LIGHT_MAX = 100
MOISTURE_MIN = 1
MOISTURE_MAX = 100
TEMP_MIN = 1
TEMP_MAX = 60

BOUNDS = {
    'light': (LIGHT_MIN, LIGHT_MAX),  # x: light intensity
    'temp': (TEMP_MIN, TEMP_MAX),    # y: temperature
    'moisture': (MOISTURE_MIN, MOISTURE_MAX)  # z: moisture
}

# === action-increment settings(5%10%20%50%70%)===
# light (1-50, 49): 5%about 2, 10%about 5, 20%about 10, 50%about 25, 70%about 35
DELTA_LIGHT = [35, 25, 10, 5, 2, -2, -5, -10, -25, -35]
# temp (1-50, 49): 5%about 2, 10%about 5, 20%about 10, 50%about 25, 70%about 35
DELTA_TEMP = [35, 25, 10, 5, 2, -2, -5, -10, -25, -35]
# moisture (1-50, 49): 5%about 2, 10%about 5, 20%about 10, 50%about 25, 70%about 35
DELTA_MOISTURE = [35, 25, 10, 5, 2, -2, -5, -10, -25, -35]


# ========================================
# ========== state normalization functions ==========
# ========================================
def normalize_state(state):
    """
    [0, 1]interval
    state: (light, temp, moisture)
    """
    normalized = np.array([
        (state[0] - BOUNDS['light'][0]) / (BOUNDS['light'][1] - BOUNDS['light'][0]),
        (state[1] - BOUNDS['temp'][0]) / (BOUNDS['temp'][1] - BOUNDS['temp'][0]),
        (state[2] - BOUNDS['moisture'][0]) / (BOUNDS['moisture'][1] - BOUNDS['moisture'][0])
    ], dtype=np.float32)
    return normalized


def denormalize_state(normalized_state):
    """
    
    normalized_state: (norm_light, norm_temp, norm_moisture) [0,1]interval
    """
    state = np.array([
        normalized_state[0] * (BOUNDS['light'][1] - BOUNDS['light'][0]) + BOUNDS['light'][0],
        normalized_state[1] * (BOUNDS['temp'][1] - BOUNDS['temp'][0]) + BOUNDS['temp'][0],
        normalized_state[2] * (BOUNDS['moisture'][1] - BOUNDS['moisture'][0]) + BOUNDS['moisture'][0]
    ])
    return state


# ========================================


def generate_random_state():
    """"""
    light = np.random.randint(BOUNDS['light'][0], BOUNDS['light'][1] + 1)
    temp = np.random.randint(BOUNDS['temp'][0], BOUNDS['temp'][1] + 1)
    moisture = np.random.randint(BOUNDS['moisture'][0], BOUNDS['moisture'][1] + 1)
    return np.array([light, temp, moisture])


def clip_state(state):
    """"""
    return np.array([
        np.clip(state[0], BOUNDS['light'][0], BOUNDS['light'][1]),
        np.clip(state[1], BOUNDS['temp'][0], BOUNDS['temp'][1]),
        np.clip(state[2], BOUNDS['moisture'][0], BOUNDS['moisture'][1])
    ])


def is_state_valid(state):
    """"""
    return (BOUNDS['light'][0] <= state[0] <= BOUNDS['light'][1] and
            BOUNDS['temp'][0] <= state[1] <= BOUNDS['temp'][1] and
            BOUNDS['moisture'][0] <= state[2] <= BOUNDS['moisture'][1])


def execute_Tr(position):
    """Execute the objective function and return triggered paths"""
    x = int(np.clip(position[0], BOUNDS['light'][0], BOUNDS['light'][1]))
    temp = int(np.clip(position[1], BOUNDS['temp'][0], BOUNDS['temp'][1]))
    z = int(np.clip(position[2], BOUNDS['moisture'][0], BOUNDS['moisture'][1]))
    return category1_multivariable_control(x, temp, z)


# === enhanced metrics collector ===
class MetricsCollector:
    def __init__(self):
        self.start_time = None
        self.end_time = None
        self.total_reward = 0
        self.td_errors = []
        self.final_output_similarities = []
        self.action_improvements = []
        self.total_memory_usage = 0
        self.memory_check_count = 0
        self.step_count = 0

        # === PSO stage statistics ===
        self.pso_start_time = None
        self.pso_end_time = None
        self.perfect_solutions_count = 0
        self.total_paths_count = 0
        self.final_fitness_scores = []
        self.pso_convergence_iterations = []
        self.pso_perfect_solutions = []
        self.pso_reset_counts = []
        self.path_execution_times = []

    def start_training(self):
        self.start_time = time.time()

    def end_training(self):
        self.end_time = time.time()

    def start_pso_phase(self):
        self.pso_start_time = time.time()

    def end_pso_phase(self):
        self.pso_end_time = time.time()

    def record_pso_result(self, fitness, is_perfect_match, convergence_iter=None, path_id=None, method='PSO',
                          reset_count=0, execution_time=0):
        self.final_fitness_scores.append(fitness)
        self.total_paths_count += 1
        self.pso_reset_counts.append(reset_count)
        self.path_execution_times.append(execution_time)

        if is_perfect_match:
            self.perfect_solutions_count += 1
            if method == 'PSO' and convergence_iter is not None:
                self.pso_perfect_solutions.append({
                    'path_id': path_id,
                    'convergence_iteration': convergence_iter,
                    'fitness': fitness,
                    'reset_count': reset_count
                })
        if convergence_iter is not None:
            self.pso_convergence_iterations.append(convergence_iter)

    def record_step_metrics(self, reward, td_error, triggered, target_path):
        self.step_count += 1
        self.total_reward += reward
        self.td_errors.append(td_error)

        try:
            process = psutil.Process(os.getpid())
            current_memory = process.memory_info().rss / 1024 / 1024
            self.total_memory_usage += current_memory
            self.memory_check_count += 1
        except:
            pass

    def record_final_output_sample(self, triggered, target_path):
        similarity = jaccard_similarity(triggered, target_path)
        self.final_output_similarities.append(similarity)

    def record_action_improvement(self, current_reward, prev_reward):
        if prev_reward is not None:
            improvement = current_reward - prev_reward
            self.action_improvements.append(1 if improvement > 0 else 0)


# === reward function ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10

    if target_path.issubset(triggered):
        reward += 1

    return reward


def category1_multivariable_control(x, y, z):
    # 确保数值在范围内
    x = max(STATE_MIN_X, min(STATE_MAX_X, x))
    y = max(STATE_MIN_Y, min(STATE_MAX_Y, y))
    z = max(STATE_MIN_Z, min(STATE_MAX_Z, z))

    triggered = set()

    if (25 < x < 75) != (25 < x < 80):
        triggered.add(1)
    if (25 < x < 75) != (30 < x < 75):
        triggered.add(2)

    if (25 < y < 75) != (40 < y < 75):
        triggered.add(3)
    if (25 < y < 75) != (50 < y < 75):
        triggered.add(4)

    if (20 < z < 50) != (30 < z < 50):
        triggered.add(5)
    if (20 < z < 50) != (20 < z < 30):
        triggered.add(6)

    if (x > 40 and x < 60 and y > 35 and y < 65) != (x > 40 and x < 75 and y > 35 and y < 65):
        triggered.add(7)
    if (x > 40 and x < 60 and y > 35 and y < 65) != (x > 40 and x < 60 and y > 20 and y < 65):
        triggered.add(8)

    if (x > 45 and x < 55 and z > 25 and z < 35) != (x > 45 and x < 55 and z > 10 and z < 35):
        triggered.add(9)

    if (y > 40 and y < 60 and z > 25 and z < 35) != (y > 40 and y < 85 and z > 25 and z < 35):
        triggered.add(10)
    if (y > 40 and y < 60 and z > 25 and z < 35) != (y > 40 and y < 60 and z > 30 and z < 40):
        triggered.add(11)

    if (x > 30 and x < 52) != (x > 30 and x < 70):
        triggered.add(12)
    if (x > 30 and x < 52) != (x > 30 and x < 80):
        triggered.add(13)

    if (y > 30 and y < 52) != (y > 30 and y < 70):
        triggered.add(14)
    if (y > 30 and y < 52) != (y > 30 and y < 80):
        triggered.add(15)

    if (z > 20 and z < 32) != (z > 20 and z < 40):
        triggered.add(16)
    if (z > 20 and z < 32) != (z > 20 and z < 50):
        triggered.add(17)

    if (abs(x - 50) < 5) != (abs(x - 50) < 10):
        triggered.add(18)
    if (abs(x - 50) < 5) != (abs(x - 30) < 5):
        triggered.add(19)

    if (abs(y - 50) < 8) != (abs(y - 50) < 15):
        triggered.add(20)
    if (abs(y - 50) < 8) != (abs(y - 20) < 8):
        triggered.add(21)

    if (abs(z - 30) < 3) != (abs(z - 30) < 6):
        triggered.add(22)
    if (abs(z - 30) < 3) != (abs(z - 20) < 3):
        triggered.add(23)

    if (x > 35 and y > 35 and z > 20) != (x > 35 and y > 60 and z > 20):
        triggered.add(24)

    if (x < 75 and y < 75 and z < 45) != (x < 75 and y < 40 and z < 45):
        triggered.add(25)
    if (x < 75 and y < 75 and z < 45) != (x < 75 and y < 30 and z < 45):
        triggered.add(26)

    if (x + y + z > 20 and x + y + z < 180) != (x + y + z > 50 and x + y + z < 180):
        triggered.add(27)
    if (x + y + z > 20 and x + y + z < 180) != (x + y + z > 60 and x + y + z < 180):
        triggered.add(28)

    if (x * y > 1000 and x * y < 4000) != (x * y > 1500 and x * y < 4000):
        triggered.add(29)
    if (x * y > 1000 and x * y < 4000) != (x * y > 1000 and x * y < 5000):
        triggered.add(30)

    if (x * z > 1200 and x * z < 3400) != (x * z > 1500 and x * z < 3400):
        triggered.add(31)
    if (x * z > 1200 and x * z < 3400) != (x * z > 1200 and x * z < 3000):
        triggered.add(32)

    if (y * z > 1000 and y * z < 3000) != (y * z > 1500 and y * z < 3000):
        triggered.add(33)
    if (y * z > 1000 and y * z < 3000) != (y * z > 1000 and y * z < 3500):
        triggered.add(34)

    if ((x + y + z) / 3 > 35 and (x + y + z) / 3 < 55) != ((x + y + z) / 3 > 40 and (x + y + z) / 3 < 55):
        triggered.add(35)
    if ((x + y + z) / 3 > 35 and (x + y + z) / 3 < 55) != ((x + y + z) / 3 > 35 and (x + y + z) / 3 < 40):
        triggered.add(36)

    if (math.sqrt(x ** 2 + y ** 2 + z ** 2) > 180) != (math.sqrt(x ** 2 + y ** 2 + z ** 2) > 100):
        triggered.add(37)
    if (math.sqrt(x ** 2 + y ** 2 + z ** 2) > 180) != (math.sqrt(x ** 2 + y ** 2 + z ** 2) > 120):
        triggered.add(38)

    if (x / y > 1 and x / y < 3) != (x / y > 1 and x / y < 1.2):
        triggered.add(39)
    if (x / y > 1 and x / y < 3) != (x / y > 1 and x / y < 2):
        triggered.add(40)

    if (x / z > 1.5 and x / z < 4.5) != (x / z > 1.8 and x / z < 4.5):
        triggered.add(41)
    if (x / z > 1.5 and x / z < 4.5) != (x / z > 1.5 and x / z < 3):
        triggered.add(42)

    if (y / z > 1.2 and y / z < 4.0) != (y / z > 1.4 and y / z < 4.0):
        triggered.add(43)
    if (y / z > 1.2 and y / z < 4.0) != (y / z > 1.2 and y / z < 3.5):
        triggered.add(44)

    if ((x - 50) + (y - 50) > 10 and (x - 50) - (y - 50) < 10) != (
            (x - 50) + (y - 50) > 20 and (x - 50) - (y - 50) < 10):
        triggered.add(45)
    if ((x - 50) + (y - 50) > 10 and (x - 50) - (y - 50) < 10) != (
            (x - 50) + (y - 50) > 50 and (x - 50) - (y - 50) < 10):
        triggered.add(46)

    if ((x - 50) + (z - 30) * 2 > 15 and (x - 50) + (z - 30) * 2 < 15) != (
            (x - 50) + (z - 30) * 2 > 10 and (x - 50) + (z - 30) * 2 < 15):
        triggered.add(47)
    if ((x - 50) + (z - 30) * 2 > 15 and (x - 50) + (z - 30) * 2 < 15) != (
            (x - 50) + (z - 30) * 2 > 15 and (x - 50) + (z - 30) * 2 < 40):
        triggered.add(48)

    if ((y - 50) + (z - 30) > 20 and (y - 50) + (z - 30) < 20) != (
            (y - 50) + (z - 30) > 10 and (y - 50) + (z - 30) < 20):
        triggered.add(49)
    if ((y - 50) + (z - 30) > 20 and (y - 50) + (z - 30) < 20) != (
            (y - 50) + (z - 30) > 20 and (y - 50) + (z - 30) < 40):
        triggered.add(50)

    if (abs((x - 50) - (y - 50)) < 8) != (abs((x - 50) - (y - 50)) < 30):
        triggered.add(51)
    if (abs((x - 50) - (y - 50)) < 8) != (abs((x - 45) - (y - 50)) < 20):
        triggered.add(52)

    if (abs((x - 50) - (z - 30) * 2) < 10) != (abs((x - 50) - (z - 30) * 2) < 30):
        triggered.add(53)
    if (abs((x - 50) - (z - 30) * 2) < 10) != (abs((x - 80) - (z - 30) * 2) < 10):
        triggered.add(54)

    if (abs((y - 50) - (z - 30)) < 5) != (abs((y - 50) - (z - 30)) < 20):
        triggered.add(55)
    if (abs((y - 50) - (z - 30)) < 5) != (abs((y - 80) - (z - 30)) < 5):
        triggered.add(56)

    if (x / (y + 10) > 0.7 and x / (y + 10) < 1.1) != (x / (y + 10) > 0.8 and x / (y + 10) < 1.5):
        triggered.add(57)
    if (x / (y + 10) > 0.7 and x / (y + 10) < 1.1) != (x / (y + 10) > 0.5 and x / (y + 10) < 1.0):
        triggered.add(58)

    if (y / (z + 20) > 1.0 and y / (z + 20) < 1.8) != (y / (z + 20) > 1.5 and y / (z + 20) < 1.8):
        triggered.add(59)

    if (z / (x / 2) > 0.8 and z / (x / 2) < 1.4) != (z / (x / 2) > 1.5 and z / (x / 2) < 1.4):
        triggered.add(60)
    if (z / (x / 2) > 0.8 and z / (x / 2) < 1.4) != (z / (x / 2) > 0.8 and z / (x / 2) < 1.9):
        triggered.add(61)

    if (x * y * z > 50000 and x * y * z < 150000) != (x * y * z > 30000 and x * y * z < 150000):
        triggered.add(62)

    if (x * 0.4 + y * 0.4 + z * 0.8 > 35 and x * 0.4 + y * 0.4 + z * 0.8 < 65) != (
            x * 0.4 + y * 0.4 + z * 0.8 > 35 and x * 0.4 + y * 0.4 + z * 0.8 < 60):
        triggered.add(63)

    if ((x * y * z) ** (1 / 3) > 25) != ((x * y * z) ** (1 / 4) > 25):
        triggered.add(64)
    if ((x * y * z) ** (1 / 3) > 25) != ((x * y * z) ** (1 / 2) > 25):
        triggered.add(65)

    if ((x - 50) * (y - 50) > 200 and (x - 50) * (y - 50) < 200) != (
            (x - 50) * (y - 50) > 150 and (x - 50) * (y - 50) < 200):
        triggered.add(66)

    if ((x - 50) * (z - 30) > 150 and (x - 50) * (z - 30) < 150) != (
            (x - 50) * (z - 30) > 100 and (x - 50) * (z - 30) < 150):
        triggered.add(67)
    if ((x - 50) * (z - 30) > 150 and (x - 50) * (z - 30) < 150) != (
            (x - 30) * (z - 30) > 150 and (x - 50) * (z - 30) < 150):
        triggered.add(68)

    if (abs(x / 50 + y / 50 + z / 30 - 4) < 0.5) != (abs(x / 50 + y / 50 + z / 30 - 4) < 1):
        triggered.add(69)
    if (abs(x / 50 + y / 50 + z / 30 - 4) < 0.5) != (abs(x / 40 + y / 50 + z / 30 - 4) < 0.5):
        triggered.add(70)

    if (abs((x / 50) * (y / 50) * (z / 30) - 1) < 0.3) != (abs((x / 50) * (y / 50) * (z / 30) - 1) < 0.8):
        triggered.add(71)
    if (abs((x / 50) * (y / 50) * (z / 30) - 1) < 0.3) != (abs((x / 80) * (y / 50) * (z / 30) - 1) < 0.3):
        triggered.add(72)

    if ((x + y) / 2 > 45 and (x + y) / 2 < 55 and z > 28 and z < 69) != (
            (x + y) / 2 > 55 and (x + y) / 2 < 55 and z > 38 and z < 69):
        triggered.add(73)
    if ((x + y) / 2 > 45 and (x + y) / 2 < 55 and z > 28 and z < 69) != (
            (x + y) / 2 > 55 and (x + y) / 2 < 53 and z > 48 and z < 69):
        triggered.add(74)

    if (abs(x - y) < 10 and z > 20 and z < 60) != (abs(x - y) < 20 and z > 20 and z < 60):
        triggered.add(75)
    if (abs(x - y) < 10 and z > 20 and z < 60) != (abs(x - y) < 10 and z > 40 and z < 60):
        triggered.add(76)

    if (math.sqrt((x - 50) ** 2 + (y - 50) ** 2) < 8 and abs(z - 30) < 3) != (
            math.sqrt((x - 50) ** 2 + (y - 50) ** 2) < 10 and abs(z - 30) < 3):
        triggered.add(77)
    if (math.sqrt((x - 50) ** 2 + (y - 50) ** 2) < 8 and abs(z - 30) < 3) != (
            math.sqrt((x - 50) ** 2 + (y - 50) ** 2) < 8 and abs(z - 30) < 10):
        triggered.add(78)

    if (x / 50 > 0.9 and x / 50 < 1.1 and y / 50 > 0.9 and y / 50 < 1.1) != (
            x / 50 > 1 and x / 50 < 1.1 and y / 50 > 0.9 and y / 50 < 1.1):
        triggered.add(79)
    if (x / 50 > 0.9 and x / 50 < 1.1 and y / 50 > 0.9 and y / 50 < 1.1) != (
            x / 50 > 1.2 and x / 50 < 1.1 and y / 50 > 0.9 and y / 50 < 1.1):
        triggered.add(80)

    if (z / 30 > 0.85 and z / 30 < 1) != (z / 30 > 0.85 and z / 30 < 1.3):
        triggered.add(81)

    if ((x / 50 + y / 50 + z / 30) / 3 > 0.9 and (x / 50 + y / 50 + z / 30) / 3 < 1) != (
            (x / 50 + y / 50 + z / 30) / 3 > 0.9 and (x / 50 + y / 50 + z / 30) / 3 < 1.5):
        triggered.add(82)

    if (max(abs(x / 50 - 1), abs(y / 50 - 1), abs(z / 30 - 1)) < 0.15) != (
            max(abs(x / 50 - 1), abs(y / 50 - 1), abs(z / 30 - 1)) < 0.25):
        triggered.add(83)
    if (max(abs(x / 50 - 1), abs(y / 50 - 1), abs(z / 30 - 1)) < 0.15) != (
            max(abs(x / 25 - 1), abs(y / 50 - 1), abs(z / 30 - 1)) < 0.15):
        triggered.add(84)

    if (min(x / 50, y / 50, z / 30) > 0.8) != (min(x / 50, y / 50, z / 30) > 1):
        triggered.add(85)
    if (min(x / 50, y / 50, z / 30) > 0.8) != (min(x / 50, y / 50, z / 30) > 1.8):
        triggered.add(86)

    if (max(x / 50, y / 50, z / 30) < 1.2) != (max(x / 50, y / 50, z / 30) < 1.6):
        triggered.add(87)

    if (abs(max(x, y, z * 2) - min(x, y, z * 2)) < 20) != (abs(max(x, y, z * 3) - min(x, y, z * 3)) < 20):
        triggered.add(88)

    if (x + y + z * 1.5 > 100 and x + y + z * 1.5 < 300) != (x + y + z * 1.5 > 180 and x + y + z * 1.5 < 300):
        triggered.add(89)

    if (x * y * z > 40000 and x * y * z < 120000) != (x * y * z > 45000 and x * y * z < 120000):
        triggered.add(90)
    if (x * y * z > 40000 and x * y * z < 120000) != (x * y * z > 40000 and x * y * z < 95000):
        triggered.add(91)

    if (abs((x + y + z) / 3 - 45) < 3) != (abs((x + y + z) / 3 - 45) < 10):
        triggered.add(92)
    if (abs((x + y + z) / 3 - 45) < 3) != (abs((x + y + z) / 3 - 30) < 3):
        triggered.add(93)

    if (abs(x * y / z - 75) < 5) != (abs(x * y / z - 75) < 10):
        triggered.add(94)

    if (x < 30 or y < 30 or z < 15) != (x < 20 or y < 30 or z < 15):
        triggered.add(95)
    if (x < 30 or y < 30 or z < 15) != (x < 30 or y < 20 or z < 15):
        triggered.add(96)

    if (x > 75 or y > 75 or z > 45) != (x > 60 or y > 75 or z > 45):
        triggered.add(97)
    if (x > 75 or y > 75 or z > 45) != (x > 75 or y > 60 or z > 45):
        triggered.add(98)

    if (abs(x / y - 1.0) > 0.25) != (abs(x / y - 1.0) > 0.3):
        triggered.add(99)
    if (abs(x / y - 1.0) > 0.25) != (abs(x / y - 0.95) > 0.25):
        triggered.add(100)

    if (abs(x / z - 1.8) > 0.3) != (abs(x / z - 1.8) > 0.4):
        triggered.add(101)

    if (abs(y / z - 1.7) > 0.3) != (abs(y / z - 1.6) > 0.3):
        triggered.add(102)

    if (x + y + z < 110 or x + y + z > 190) != (x + y + z < 115 or x + y + z > 190):
        triggered.add(103)

    if (x * y * z < 60000 or x * y * z > 140000) != (x * y * z < 60000 or x * y * z > 135000):
        triggered.add(104)

    if (x < 35 and y < 35) != (x < 30 and y < 35):
        triggered.add(105)
    if (x < 35 and y < 35) != (x < 35 and y < 30):
        triggered.add(106)

    if (y < 35 and z < 20) != (y < 30 and z < 20):
        triggered.add(107)

    if (x < 25 and y < 25 and z < 15) != (x < 22 and y < 25 and z < 15):
        triggered.add(108)
    if (x < 25 and y < 25 and z < 15) != (x < 25 and y < 22 and z < 15):
        triggered.add(109)

    return triggered


# target path definitions
targetPaths = [
    {4, 6, 11, 12, 13, 16, 17, 18, 21, 24, 25, 26, 36, 39, 41, 43, 48, 51, 52, 54, 55, 58, 60, 64, 68, 69, 70, 71, 72,
     73, 74, 75, 81, 82, 83, 85, 86, 87, 89, 91, 93, 99, 102},
    {6, 10, 12, 13, 14, 15, 16, 17, 18, 20, 22, 25, 26, 36, 41, 45, 46, 47, 49, 51, 52, 54, 55, 57, 59, 60, 64, 69, 70,
     71, 72, 76, 81, 82, 86, 87, 88, 89, 91, 92, 98, 102},
    {6, 7, 11, 12, 13, 14, 15, 16, 17, 21, 24, 25, 26, 36, 39, 41, 43, 48, 49, 51, 52, 54, 55, 58, 60, 64, 67, 68, 72,
     75, 82, 86, 87, 89, 92, 97, 100, 102, 104},
    {6, 11, 12, 13, 14, 15, 16, 17, 19, 21, 24, 25, 26, 36, 48, 49, 53, 56, 57, 61, 64, 68, 69, 70, 71, 72, 73, 74, 76,
     78, 81, 82, 83, 86, 87, 89, 91, 92, 101},
    {6, 7, 13, 14, 15, 16, 17, 20, 22, 24, 25, 26, 30, 36, 39, 48, 49, 51, 52, 53, 56, 58, 59, 60, 64, 66, 68, 72, 75,
     81, 82, 86, 87, 88, 89, 92, 97, 100, 104},
    {4, 6, 12, 13, 16, 17, 19, 21, 22, 24, 25, 26, 33, 36, 39, 41, 43, 47, 51, 52, 54, 55, 60, 64, 69, 72, 73, 74, 76,
     78, 81, 82, 84, 85, 86, 89, 93, 94, 100},
    {5, 7, 11, 12, 13, 14, 15, 21, 22, 24, 25, 26, 33, 36, 39, 47, 51, 52, 53, 54, 55, 58, 59, 64, 69, 70, 71, 72, 75,
     82, 85, 86, 87, 88, 89, 91, 92, 97, 100},
    {6, 10, 14, 15, 16, 17, 18, 20, 22, 24, 25, 26, 31, 36, 49, 51, 52, 53, 55, 58, 59, 61, 64, 69, 73, 74, 75, 81, 82,
     83, 85, 86, 87, 89, 93, 94, 99, 100},
    {5, 7, 11, 13, 14, 15, 21, 23, 24, 25, 26, 30, 36, 39, 48, 51, 53, 54, 55, 58, 59, 60, 64, 66, 71, 72, 75, 82, 85,
     86, 87, 88, 89, 91, 92, 97, 99, 102},
    {6, 10, 12, 13, 15, 16, 17, 19, 23, 25, 26, 36, 41, 46, 50, 51, 52, 54, 55, 58, 59, 60, 64, 71, 72, 75, 81, 82, 86,
     87, 88, 89, 91, 92, 98, 99, 100},
    {4, 6, 11, 12, 13, 16, 17, 19, 21, 22, 24, 25, 26, 36, 43, 47, 54, 55, 60, 64, 69, 72, 73, 74, 76, 78, 80, 81, 82,
     83, 85, 86, 89, 93, 94, 101, 102},
    {5, 6, 11, 14, 15, 19, 20, 23, 24, 25, 26, 31, 36, 41, 51, 52, 54, 55, 57, 59, 60, 64, 69, 72, 73, 74, 75, 77, 81,
     82, 83, 85, 86, 88, 89, 93, 102},
    {5, 6, 11, 14, 15, 19, 21, 23, 24, 25, 26, 31, 36, 41, 54, 56, 57, 59, 60, 63, 64, 69, 72, 73, 74, 76, 79, 80, 81,
     82, 84, 85, 86, 88, 89, 93, 94},
    {5, 6, 7, 10, 13, 14, 15, 20, 23, 25, 26, 30, 37, 46, 48, 49, 51, 52, 53, 54, 55, 59, 60, 64, 72, 76, 81, 82, 85,
     86, 87, 88, 89, 97, 98, 104},
    {5, 9, 12, 13, 14, 15, 19, 20, 25, 26, 31, 33, 36, 45, 46, 51, 52, 53, 55, 57, 59, 60, 63, 64, 69, 72, 75, 82, 83,
     87, 88, 89, 93, 98, 101},
    {3, 4, 6, 7, 12, 13, 16, 17, 20, 22, 24, 26, 33, 36, 39, 48, 51, 54, 55, 57, 60, 64, 68, 69, 70, 72, 73, 74, 81, 82,
     87, 89, 93, 94, 97},
    {1, 5, 10, 13, 14, 15, 20, 22, 30, 36, 37, 39, 48, 51, 52, 54, 55, 58, 59, 64, 71, 72, 75, 82, 85, 86, 87, 88, 89,
     92, 97, 98, 100},
    {5, 7, 13, 14, 15, 20, 25, 26, 30, 33, 36, 42, 47, 51, 52, 54, 55, 58, 59, 64, 69, 70, 71, 72, 75, 82, 87, 88, 89,
     91, 92, 97, 98},
    {3, 4, 6, 16, 17, 19, 20, 23, 24, 26, 33, 35, 36, 39, 41, 43, 51, 52, 54, 55, 58, 60, 63, 64, 72, 75, 81, 83, 89,
     92, 99},
    {12, 13, 15, 23, 25, 26, 30, 31, 33, 36, 37, 42, 44, 46, 49, 54, 56, 64, 69, 70, 71, 72, 82, 87, 88, 89, 91, 92, 97,
     98},
    {6, 7, 13, 14, 15, 17, 20, 24, 30, 32, 37, 39, 41, 43, 50, 51, 52, 54, 55, 58, 60, 64, 70, 75, 82, 86, 87, 97, 100,
     102},
    {6, 12, 13, 15, 17, 18, 30, 34, 37, 46, 48, 50, 51, 52, 53, 55, 58, 59, 61, 64, 66, 68, 82, 86, 87, 99, 100, 101,
     102},
    {3, 4, 6, 13, 16, 17, 22, 26, 33, 36, 39, 40, 48, 53, 60, 64, 68, 69, 70, 72, 73, 74, 81, 82, 87, 89, 93, 97, 101},
    {3, 4, 6, 8, 12, 13, 16, 17, 19, 21, 22, 29, 35, 39, 40, 47, 51, 54, 57, 60, 62, 64, 71, 81, 89, 92, 96, 101, 103},
    {2, 6, 10, 14, 15, 16, 17, 19, 22, 25, 26, 36, 50, 53, 55, 59, 63, 64, 72, 73, 74, 81, 82, 87, 89, 93, 95, 98},
    {3, 4, 5, 9, 12, 13, 19, 20, 23, 24, 26, 35, 39, 51, 52, 53, 56, 58, 60, 62, 64, 71, 75, 88, 89, 90, 92, 103},
    {6, 15, 16, 17, 37, 38, 39, 50, 51, 53, 54, 55, 58, 59, 60, 64, 69, 75, 81, 86, 88, 100, 102},
    {3, 4, 31, 35, 48, 49, 51, 52, 58, 62, 64, 69, 71, 82, 89, 90, 92, 95, 99, 100, 103, 106},
    {3, 4, 12, 13, 23, 26, 35, 39, 42, 51, 54, 55, 57, 62, 64, 71, 87, 89, 92, 103, 107},
    {3, 4, 6, 17, 19, 26, 29, 31, 33, 35, 47, 62, 63, 64, 71, 87, 89, 92, 103, 105, 106},
    {7, 12, 13, 14, 15, 21, 25, 26, 35, 39, 51, 52, 56, 58, 65, 87, 89, 92, 97, 100},
    {3, 4, 26, 28, 51, 52, 53, 54, 55, 58, 59, 65, 99, 100, 106, 107},
    {3, 4, 26, 27, 28, 41, 51, 52, 54, 55, 59, 60, 65, 106, 107},
    {27, 28, 39, 40, 42, 51, 52, 53, 54, 55, 57, 65, 108},
    {21, 27, 28, 53, 54, 56, 58, 59, 65, 88, 109},
]


def jaccard_similarity(set1, set2):
    """
    Compute Jaccard similarity
    **If set1 contains set2(set2set1), return 1.0**
    """
    # set1set2(target paths), Similarity1
    if set2.issubset(set1):
        return 1.0

    # Otherwise compute standard Jaccard similarity
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    return intersection / union if union != 0 else 0.0


def calculate_fitness(position, target_path):
    """(JaccardSimilarity)"""
    triggered = execute_Tr(position)

    # Path target paths, Maximum
    if target_path.issubset(triggered):
        return 1.0

    # Compute Jaccard similarity
    intersection = len(triggered & target_path)
    union = len(triggered | target_path)
    return intersection / union if union > 0 else 0.0


# === Path Similarity ===
def compute_path_similarity_matrix(paths):
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix


# === Path  ===
def group_paths_by_similarity(paths):
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.mean(avg_sim_scores)

    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]
    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)

    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group


# === Sample generation===
def generate_samples_for_similar_paths(similar_group_indices, num_total=2000, top_k=200):
    def jaccard_similarity_local(a, b):
        if not a and not b:
            return 1.0
        return len(a & b) / len(a | b) if a | b else 0.0

    def compute_robustness(state, path):
        base = execute_Tr(state)
        if not base:
            return 0.0
        rob, neighbors = 0.0, 0
        for dx in [-1, 0, 1]:
            for dy in [-1, 0, 1]:
                for dz in [-1, 0, 1]:
                    if dx == dy == dz == 0:
                        continue
                    neighbor = clip_state(state + np.array([dx, dy, dz]))
                    n_trig = execute_Tr(neighbor)
                    if not n_trig:
                        continue
                    rob += jaccard_similarity_local(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Path {path_id}\n")
            f.write("light temp moisture\tScore\tSimilarity\tLengthDiff\tRobustness\n")
            for s in samples:
                light, temp, moisture = int(s[0][0]), int(s[0][1]), int(s[0][2])
                f.write(f"{light} {temp} {moisture}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\n")

    print("Similar path group...")
    base_dir = os.path.join(os.getcwd(), "path_samples")
    for path_idx in similar_group_indices:
        path = targetPaths[path_idx]
        samples = []
        attempts = 0
        while len(samples) < top_k and attempts < num_total * 5:
            attempts += 1
            state = generate_random_state()
            triggered = execute_Tr(state)
            if not triggered:
                continue
            sim = jaccard_similarity_local(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            score = 0.55 * sim + 0.25 * len_diff + 0.2 * rob
            samples.append((state, score, sim, len_diff, rob))
        if samples:
            samples.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples[:top_k], base_dir=base_dir)


# === ===
class SharedExperienceReplay:
    def __init__(self, capacity=10000):
        self.capacity = capacity
        self.buffer = deque(maxlen=self.capacity)
        self.priorities = deque(maxlen=self.capacity)

    def append(self, experience):
        self.buffer.append(experience)
        self.priorities.append(experience[-1])

    def sample(self, batch_size, alpha=0.6):
        if len(self.buffer) == 0:
            return [], [], []

        priorities = np.array(self.priorities) ** alpha
        sum_priorities = np.sum(priorities)
        if sum_priorities == 0:
            probabilities = np.ones(len(self.buffer)) / len(self.buffer)
        else:
            probabilities = priorities / sum_priorities

        batch_indices = np.random.choice(len(self.buffer), batch_size, p=probabilities)
        batch = [self.buffer[idx] for idx in batch_indices]
        return batch, batch_indices, probabilities[batch_indices]

    def __len__(self):
        return len(self.buffer)

    def get_high_reward_samples(self, target_path, num_samples=20):
        """target pathsSimilarityMaximum"""
        if len(self.buffer) == 0:
            return []

        samples_with_similarity = []
        for experience in self.buffer:
            state_tensor = experience[0]
            state_tuple = tuple(state_tensor.cpu().numpy().flatten().astype(int))
            triggered = execute_Tr(state_tuple)
            sim = jaccard_similarity(triggered, target_path)

            # (Similarity1), 
            if sim >= 1.0:
                return [(state_tuple, 0, sim, triggered)]

            samples_with_similarity.append((state_tuple, 0, sim, triggered))

        # Similarity, num_samples
        samples_with_similarity.sort(key=lambda x: x[2], reverse=True)
        return samples_with_similarity[:num_samples]


def load_path_data(file_path):
    path_data = []
    if not os.path.exists(file_path):
        return path_data

    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines[2:]:
            parts = line.strip().split("\t")
            if parts:
                state = tuple(map(int, parts[0].split()))
                path_data.append(state)
    return path_data


# === DQN ===
class DQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, 128)
        self.fc2 = nn.Linear(128, 64)
        self.fc3 = nn.Linear(64, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)


# === DQN Agent===
class DQNAgentWithPER:
    def __init__(self, state_dim, action_dim, replay_buffer, gamma=0.99, epsilon=1.0, epsilon_decay=0.995,
                 epsilon_min=0.1, learning_rate=0.001, alpha=0.6, beta=0.4):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer
        self.alpha = alpha
        self.beta = beta

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        """"""
        dim = action_idx // 10  # : 0=light, 1=temp, 2=moisture
        delta_idx = action_idx % 10  # : 0-9

        if dim == 0:  # light
            delta = DELTA_LIGHT[delta_idx]
            return (delta, 0, 0)
        elif dim == 1:  # temp
            delta = DELTA_TEMP[delta_idx]
            return (0, delta, 0)
        elif dim == 2:  # moisture
            delta = DELTA_MOISTURE[delta_idx]
            return (0, 0, delta)

    def act(self, state):
        if random.random() < self.epsilon:
            return random.randrange(self.action_dim)
        state = torch.tensor(state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state)
        return torch.argmax(q_values, dim=1).item()

    def store_transition(self, state, action, reward, next_state, done):
        state = torch.tensor(state, dtype=torch.float32).unsqueeze(0).to(device)
        next_state = torch.tensor(next_state, dtype=torch.float32).unsqueeze(0).to(device)

        with torch.no_grad():
            q_values = self.model(state)
            next_q_values = self.target_model(next_state)
            max_next_q_values = next_q_values.max(1)[0]
            target_q_values = reward + (self.gamma * max_next_q_values * (1 - done))
            td_error = torch.abs(q_values[0][action] - target_q_values).item()

        self.replay_buffer.append((state, action, reward, next_state, done, td_error))
        return td_error

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return

        batch, batch_indices, probabilities = self.replay_buffer.sample(batch_size, alpha=self.alpha)
        states, actions, rewards, next_states, dones, _ = zip(*batch)

        states = torch.tensor(np.array([s.cpu().numpy().flatten() for s in states]), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array([ns.cpu().numpy().flatten() for ns in next_states]),
                                   dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))

        loss = nn.MSELoss()(current_q_values, target_q_values)

        self.optimizer.zero_grad()
        loss.backward()
        self.optimizer.step()

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())


# === Run ===
def generate_and_train_for_similar_paths(agent, similar_group, path_documents, run_metrics, episodes=500, batch_size=32,
                                         steps_per_test=5, replay_times=10, is_isolated=False):
    trained_paths = set()

    for episode in range(episodes):
        for path_idx in similar_group:
            if path_idx in trained_paths:
                continue

            file_path = os.path.join(path_documents, f"path{path_idx + 1}{'_isolated' if is_isolated else ''}.txt")
            path_data = load_path_data(file_path)
            if not path_data:
                trained_paths.add(path_idx)
                continue

            target_path = targetPaths[path_idx]

            # ===  ===
            BATCH_SIZE = 50  # 50
            N_SAMPLES = 200  # 200
            N_STEPS = 3  # 3
            N_EPOCHS = 5  # 5

            replay_count = 0  # , 

            # 5
            for epoch in range(N_EPOCHS):
                print(f"  Path {path_idx + 1} - Run {epoch + 1}/{N_EPOCHS}")

                # 2004, 50
                for batch_start in range(0, N_SAMPLES, BATCH_SIZE):
                    batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)

                    # 50
                    for test_data_idx in range(batch_start, batch_end):
                        if test_data_idx >= len(path_data):
                            break

                        state = path_data[test_data_idx]
                        prev_state = None
                        prev_triggered = None
                        prev_reward = None

                        # 3
                        for step in range(N_STEPS):
                            # 
                            legal_actions = []
                            for a in range(agent.action_dim):
                                dx, dy, dz = agent.decode_action(a)
                                cand_next = (state[0] + dx, state[1] + dy, state[2] + dz)
                                if is_state_valid(cand_next):
                                    legal_actions.append(a)

                            if not legal_actions:
                                break

                            # 
                            if random.random() < agent.epsilon:
                                action = random.choice(legal_actions)
                            else:
                                state_tensor = torch.tensor(state, dtype=torch.float32).unsqueeze(0).to(device)
                                with torch.no_grad():
                                    q_values = agent.model(state_tensor)[0]
                                action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                            # 
                            dx, dy, dz = agent.decode_action(action)
                            next_state = (state[0] + dx, state[1] + dy, state[2] + dz)

                            # 
                            triggered = execute_Tr(next_state)
                            reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                            done = (step == N_STEPS - 1)

                            # 
                            td_error = agent.store_transition(state, action, reward, next_state, done)
                            run_metrics.record_step_metrics(reward, td_error, triggered, target_path)

                            if prev_reward is not None:
                                run_metrics.record_action_improvement(reward, prev_reward)

                            # 
                            prev_state = state
                            prev_triggered = triggered
                            prev_reward = reward
                            state = next_state

                    # 50, 
                    if len(agent.replay_buffer) >= batch_size:
                        agent.train(batch_size)
                        replay_count += 1

                        # 2
                        if replay_count % 2 == 0:
                            agent.update_target_model()
                            print(
                                f"     {batch_start // BATCH_SIZE + 1}/4 completed |  {replay_count}  | ")

            trained_paths.add(path_idx)
            print(f"  Path {path_idx + 1} completed\n")

        if len(trained_paths) == len(similar_group):
            break

    return agent


def generate_samples_for_isolated_paths(agent_similar, isolated_group_indices, num_total=2000, top_k=200):
    def compute_q_value(state, agent):
        state_tensor = torch.tensor(state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = agent.model(state_tensor)
        return q_values.max().item()

    def compute_robustness(state, path):
        base = execute_Tr(state)
        if not base:
            return 0.0
        rob, neighbors = 0.0, 0
        for dx in [-1, 0, 1]:
            for dy in [-1, 0, 1]:
                for dz in [-1, 0, 1]:
                    if dx == dy == dz == 0:
                        continue
                    neighbor = clip_state(state + np.array([dx, dy, dz]))
                    n_trig = execute_Tr(neighbor)
                    if not n_trig:
                        continue
                    rob += jaccard_similarity(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_isolated.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Isolated Path {path_id}\n")
            f.write("light temp moisture\tScore\tSimilarity\tLengthDiff\tRobustness\tQ_value_normalized_complement\n")
            for s in samples:
                light, temp, moisture = int(s[0][0]), int(s[0][1]), int(s[0][2])
                f.write(f"{light} {temp} {moisture}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\t{s[5]:.4f}\n")

    base_dir = os.path.join(os.getcwd(), "path_samples")

    for path_idx in isolated_group_indices:
        path = targetPaths[path_idx]
        candidate_samples = []
        attempts = 0

        while len(candidate_samples) < num_total and attempts < num_total * 5:
            attempts += 1
            state = generate_random_state()
            triggered = execute_Tr(state)
            if not triggered:
                continue

            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            q_value = compute_q_value(state, agent_similar)

            candidate_samples.append((state, sim, len_diff, rob, q_value))

        if not candidate_samples:
            continue

        q_values = [sample[4] for sample in candidate_samples]
        q_min = min(q_values)
        q_max = max(q_values)

        normalized_samples = []
        for state, sim, len_diff, rob, q_value in candidate_samples:
            if q_max - q_min > 0:
                q_normalized = (q_value - q_min) / (q_max - q_min)
            else:
                q_normalized = 0.5

            q_complement = 1.0 - q_normalized
            score = 0.28 * sim + 0.1 * len_diff + 0.19 * rob + 0.43 * q_complement

            normalized_samples.append((state, score, sim, len_diff, rob, q_complement))

        normalized_samples.sort(key=lambda x: x[1], reverse=True)
        top_samples = normalized_samples[:top_k]

        save_samples(path_id=path_idx + 1, samples=top_samples, base_dir=base_dir)


# === Run ===
def generate_and_train_for_isolated_paths_enhanced(agent_similar, agent_isolated, similar_group, isolated_group,
                                                   path_documents, run_metrics, episodes=500, batch_size=32,
                                                   is_isolated=True):
    trained_paths = set()

    for episode in range(episodes):
        for path_idx in isolated_group:
            if path_idx in trained_paths:
                continue

            file_path = os.path.join(path_documents, f"path{path_idx + 1}_isolated.txt")
            stage2_path_data = load_path_data(file_path)
            if not stage2_path_data:
                trained_paths.add(path_idx)
                continue

            target_path = targetPaths[path_idx]

            # ===  ===
            BATCH_SIZE = 50  # 50
            N_SAMPLES = 200  # 200
            N_STEPS = 3  # 3
            N_EPOCHS = 5  # 5

            replay_count = 0  # 

            # 5
            for epoch in range(N_EPOCHS):
                print(f"  Path {path_idx + 1} - Run {epoch + 1}/{N_EPOCHS}")

                # 2004,50
                for batch_start in range(0, N_SAMPLES, BATCH_SIZE):
                    batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)

                    # 50
                    for test_data_idx in range(batch_start, batch_end):
                        if test_data_idx >= len(stage2_path_data):
                            break

                        state = stage2_path_data[test_data_idx]
                        prev_state = None
                        prev_triggered = None
                        prev_reward = None

                        # 3
                        for step in range(N_STEPS):
                            # 
                            legal_actions = []
                            for a in range(agent_isolated.action_dim):
                                dx, dy, dz = agent_isolated.decode_action(a)
                                cand_next = (state[0] + dx, state[1] + dy, state[2] + dz)
                                if is_state_valid(cand_next):
                                    legal_actions.append(a)

                            if not legal_actions:
                                break

                            # 
                            if random.random() < agent_isolated.epsilon:
                                action = random.choice(legal_actions)
                            else:
                                state_tensor = torch.tensor(state, dtype=torch.float32).unsqueeze(0).to(device)
                                with torch.no_grad():
                                    q_values = agent_isolated.model(state_tensor)[0]
                                action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                            # 
                            dx, dy, dz = agent_isolated.decode_action(action)
                            next_state = (state[0] + dx, state[1] + dy, state[2] + dz)

                            # 
                            triggered = execute_Tr(next_state)
                            reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)

                            # Path 
                            if target_path.issubset(triggered):
                                reward += 2.0

                            done = (step == N_STEPS - 1)

                            # 
                            td_error = agent_isolated.store_transition(state, action, reward, next_state, done)
                            run_metrics.record_step_metrics(reward, td_error, triggered, target_path)

                            if prev_reward is not None:
                                run_metrics.record_action_improvement(reward, prev_reward)

                            # 
                            prev_state = state
                            prev_triggered = triggered
                            prev_reward = reward
                            state = next_state

                    # 50, 
                    if len(agent_isolated.replay_buffer) >= batch_size:
                        agent_isolated.train(batch_size)
                        replay_count += 1

                        # 2
                        if replay_count % 2 == 0:
                            agent_isolated.update_target_model()
                            print(
                                f"     {batch_start // BATCH_SIZE + 1}/4 completed |  {replay_count}  | ")

            trained_paths.add(path_idx)
            print(f"  Path {path_idx + 1} completed\n")

        if len(trained_paths) == len(isolated_group):
            break

    return agent_isolated


# === PSO ===
class Particle:
    def __init__(self, initial_position=None):
        if initial_position is not None:
            self.position = np.array(initial_position, dtype=float)
        else:
            self.position = np.array([
                np.random.uniform(BOUNDS['light'][0], BOUNDS['light'][1]),
                np.random.uniform(BOUNDS['temp'][0], BOUNDS['temp'][1]),
                np.random.uniform(BOUNDS['moisture'][0], BOUNDS['moisture'][1])
            ])

        self.velocity = np.array([
            np.random.uniform(-5, 5),
            np.random.uniform(-3, 3),
            np.random.uniform(-5, 5)
        ])

        self.best_position = self.position.copy()
        self.best_fitness = 0
        self.fitness = 0


# === standard PSO ===
class PSO:
    def __init__(self, target_path, swarm_size=20, dqn_samples=None):
        self.target_path = target_path
        self.swarm_size = swarm_size
        self.particles = []
        self.global_best_position = None
        self.global_best_fitness = 0
        self.reset_count = 0

        if dqn_samples is not None and len(dqn_samples) > 0:
            num_direct = min(len(dqn_samples), swarm_size)
            for i in range(num_direct):
                state_tuple, reward, sim, triggered = dqn_samples[i]
                particle = Particle(initial_position=state_tuple)
                self.particles.append(particle)

            if len(self.particles) < swarm_size:
                remaining = swarm_size - len(self.particles)
                for i in range(remaining):
                    base_idx = i % len(dqn_samples)
                    state_tuple, _, _, _ = dqn_samples[base_idx]
                    perturbed = np.array(state_tuple) + np.random.randint(-10, 11, size=3)
                    perturbed = clip_state(perturbed)
                    particle = Particle(initial_position=perturbed.tolist())
                    self.particles.append(particle)
        else:
            self.particles = [Particle() for _ in range(swarm_size)]

        for particle in self.particles:
            particle.fitness = self.fitness_function(particle.position)
            if particle.fitness > particle.best_fitness:
                particle.best_fitness = particle.fitness
                particle.best_position = particle.position.copy()
            if particle.fitness > self.global_best_fitness:
                self.global_best_fitness = particle.fitness
                self.global_best_position = particle.position.copy()

    def fitness_function(self, position):
        try:
            triggered = execute_Tr(position)
            # : target pathsPath , 1
            if self.target_path.issubset(triggered):
                return 1.0

            # Compute Jaccard similarity
            intersection = len(triggered & self.target_path)
            union = len(triggered | self.target_path)
            return intersection / union if union > 0 else 0.0
        except:
            return 0.0

    def update(self, iteration, max_iterations):
        # PSO
        w = 0.7  # 
        c1 = 1.5  # 
        c2 = 1.5  # 

        for particle in self.particles:
            r1 = np.random.random(3)
            r2 = np.random.random(3)

            # 
            particle.velocity = (w * particle.velocity +
                                 c1 * r1 * (particle.best_position - particle.position) +
                                 c2 * r2 * (self.global_best_position - particle.position))

            # 
            max_velocity = np.array([
                (BOUNDS['light'][1] - BOUNDS['light'][0]) * 0.2,
                (BOUNDS['temp'][1] - BOUNDS['temp'][0]) * 0.2,
                (BOUNDS['moisture'][1] - BOUNDS['moisture'][0]) * 0.2
            ])
            particle.velocity = np.clip(particle.velocity, -max_velocity, max_velocity)

            # 
            particle.position += particle.velocity
            particle.position = clip_state(particle.position)

            # 
            particle.fitness = self.fitness_function(particle.position)

            # 
            if particle.fitness > particle.best_fitness:
                particle.best_fitness = particle.fitness
                particle.best_position = particle.position.copy()

            # 
            if particle.fitness > self.global_best_fitness:
                self.global_best_fitness = particle.fitness
                self.global_best_position = particle.position.copy()


# === Excel ===
def export_multiple_runs_to_excel(all_run_results, all_run_metrics, num_runs, filename=None):
    """ runExcel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"DQN_PSO_{num_runs}Runs_{timestamp}.xlsx"

    wb = Workbook()

    # 
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")

    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    dqn_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # ========== 1:  ==========
    ws1 = wb.active
    ws1.title = ""
    ws1.sheet_view.showGridLines = False

    # 
    headers = ["Run", "", "", "", "Average Iterations", "(s)", "DQN"]
    col_widths = [12, 12, 12, 14, 14, 14, 12]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    # 
    for run_idx, (results, run_metrics) in enumerate(zip(all_run_results, all_run_metrics), start=1):
        success_count = sum(1 for r in results if r['perfect_match'])
        success_rate = (success_count / len(targetPaths)) * 100
        avg_fitness = np.mean([r['fitness'] for r in results])

        # Average Iterations
        iterations_list = []
        for r in results:
            if r.get('method') == 'DQN':
                iterations_list.append(0)
            elif r.get('convergence_iteration') is not None:
                iterations_list.append(r['convergence_iteration'])
            else:
                iterations_list.append(10000)
        avg_iterations = np.mean(iterations_list)

        # DQN
        dqn_solved_count = sum(1 for r in results if r.get('method') == 'DQN')

        # 
        total_time = run_metrics.pso_end_time - run_metrics.pso_start_time if run_metrics.pso_end_time else 0

        row_data = [
            f" {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(targetPaths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}",
            f"{dqn_solved_count}/{len(targetPaths)}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

            if col == 7 and dqn_solved_count > 0:
                cell.fill = dqn_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:G{len(all_run_results) + 1}"

    # ========== 2: Path  ==========
    ws2 = wb.create_sheet(title="Path ")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Path ID", "", "", "", "Average Iterations", "Minimum Iterations", "Maximum Iterations",
                "DQN"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    num_paths = len(targetPaths)
    for path_idx in range(num_paths):
        success_count = sum(1 for results in all_run_results if results[path_idx]['perfect_match'])
        success_rate = (success_count / num_runs) * 100
        avg_fitness = np.mean([results[path_idx]['fitness'] for results in all_run_results])

        iterations_list = []
        for results in all_run_results:
            r = results[path_idx]
            if r.get('method') == 'DQN':
                iterations_list.append(0)
            elif r.get('convergence_iteration') is not None:
                iterations_list.append(r['convergence_iteration'])
            else:
                iterations_list.append(10000)

        avg_iterations = np.mean(iterations_list) if iterations_list else 0
        min_iterations = np.min(iterations_list) if iterations_list else 0
        max_iterations = np.max(iterations_list) if iterations_list else 0

        dqn_solved_count = sum(1 for results in all_run_results
                               if results[path_idx].get('method') == 'DQN')

        row_data = [
            f"Path  {path_idx + 1}",
            f"{success_count}/{num_runs}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}",
            f"{dqn_solved_count}/{num_runs}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

            if col == 8 and dqn_solved_count > 0:
                cell.fill = dqn_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:H{len(targetPaths) + 1}"

    # ========== 3:  ==========
    ws3 = wb.create_sheet(title="")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Path ", "", "(light,temp,moisture)", "", "Iterations", "", "Path "]
    col_widths3 = [10, 10, 25, 12, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(num_paths):
        for run_idx, results in enumerate(all_run_results, start=1):
            result = results[path_idx]
            best_position = result['best_position']
            fitness = result['fitness']
            triggered = result['triggered']
            method = result.get('method', 'PSO')

            if method == 'DQN':
                convergence_iter = 0
            elif result.get('convergence_iteration') is not None:
                convergence_iter = result['convergence_iteration']
            else:
                convergence_iter = 10000

            particle_str = f"({int(best_position[0])}, {int(best_position[1])}, {int(best_position[2])})"
            path_str = str(sorted(list(triggered)))

            row_data = [
                f"Path {path_idx + 1}",
                f"{run_idx}",
                particle_str,
                f"{fitness:.4f}",
                convergence_iter if convergence_iter < 10000 else "-",
                method,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 7:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if fitness == 1.0:
                    if method == 'DQN':
                        cell.fill = dqn_fill
                    else:
                        cell.fill = success_fill
                elif fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:G{row_idx - 1}"

    # ========== 4: target paths ==========
    ws4 = wb.create_sheet(title="target paths")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Path ID", "target paths", ""]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(targetPaths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"Path  {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    # 
    output_dir = os.path.join(os.getcwd(), "results")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)

    print(f"\n{'=' * 70}")
    print(f" : {filepath}")
    print(f"{'=' * 70}")
    print(f":")
    print(f"  1.        - {num_runs} run(DQN)")
    print(f"  2. Path        - Path (DQN)")
    print(f"  3.    -  runPath ()")
    print(f"  4. target paths       - target paths")
    print(f"{'=' * 70}\n")

    return filepath


def run_single_experiment(run_num, similar_group, isolated_group):
    """Run oneDQN-PSO"""
    print(f"\n{'=' * 100}")
    print(f"Start run  {run_num}  run")
    print(f"{'=' * 100}")

    run_metrics = MetricsCollector()
    run_metrics.start_training()

    path_documents = os.path.join(os.getcwd(), "path_samples")

    if run_num == 1:
        print("Path ...")
        generate_samples_for_similar_paths(similar_group, num_total=2000, top_k=200)

    replay_buffer = SharedExperienceReplay(capacity=10000)
    state_dim = 3
    action_dim = 30
    agent = DQNAgentWithPER(state_dim, action_dim, replay_buffer)

    print(f"{run_num} - Run : Path ")
    generate_and_train_for_similar_paths(agent, similar_group, path_documents, run_metrics, episodes=500, batch_size=32,
                                         is_isolated=False)

    if run_num == 1:
        print("Path ...")
        generate_samples_for_isolated_paths(agent, isolated_group, num_total=2000, top_k=200)

    print(f"{run_num} - Run : Path ")
    isolated_replay_buffer = SharedExperienceReplay(capacity=15000)
    agent_isolated = DQNAgentWithPER(state_dim, action_dim, isolated_replay_buffer)

    agent_isolated.model.load_state_dict(agent.model.state_dict())
    agent_isolated.target_model.load_state_dict(agent.model.state_dict())

    agent_isolated = generate_and_train_for_isolated_paths_enhanced(
        agent_similar=agent,
        agent_isolated=agent_isolated,
        similar_group=similar_group,
        isolated_group=isolated_group,
        path_documents=path_documents,
        run_metrics=run_metrics,
        episodes=500,
        batch_size=32,
        is_isolated=True
    )

    run_metrics.end_training()

    # final samples
    dqn_best_samples = {}

    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        dqn_best_samples[path_idx] = high_reward_samples
        for state_tuple, _, sim, triggered in high_reward_samples:
            run_metrics.record_final_output_sample(triggered, target_path)

    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        dqn_best_samples[path_idx] = high_reward_samples
        for state_tuple, _, sim, triggered in high_reward_samples:
            run_metrics.record_final_output_sample(triggered, target_path)

    # PSO
    print(f"{run_num} - PSO")
    run_metrics.start_pso_phase()

    max_iterations = 3000
    pso_results = []

    for i, target_path in enumerate(targetPaths):
        path_start_time = time.time()
        dqn_samples_for_path = dqn_best_samples.get(i, [])

        perfect_solution_found = False
        perfect_solution_state = None

        if dqn_samples_for_path:
            for sample in dqn_samples_for_path:
                state_tuple, reward, sim, triggered = sample
                if target_path.issubset(triggered):
                    perfect_solution_found = True
                    perfect_solution_state = state_tuple
                    break

        if perfect_solution_found:
            path_execution_time = time.time() - path_start_time
            pso_results.append({
                'target_path': target_path,
                'best_position': np.array(perfect_solution_state),
                'fitness': 1.0,
                'triggered': triggered,
                'perfect_match': True,
                'method': 'DQN',
                'convergence_iteration': 0,
                'early_stopped': False,
                'reset_count': 0
            })
            run_metrics.record_pso_result(1.0, True, convergence_iter=0, path_id=i + 1,
                                          method='DQN', reset_count=0, execution_time=path_execution_time)
            status = "(DQN)"
        else:
            pso = PSO(target_path, swarm_size=20, dqn_samples=dqn_samples_for_path)

            converged_at_iteration = max_iterations
            early_stop = False

            for iteration in range(max_iterations):
                pso.update(iteration, max_iterations)

                if pso.global_best_fitness >= 1.0:
                    converged_at_iteration = iteration + 1
                    early_stop = True
                    break

            path_execution_time = time.time() - path_start_time
            best_position = pso.global_best_position
            triggered = execute_Tr(best_position)

            is_perfect = target_path.issubset(triggered)

            pso_results.append({
                'target_path': target_path,
                'best_position': best_position,
                'fitness': pso.global_best_fitness,
                'triggered': triggered,
                'perfect_match': is_perfect,
                'method': 'PSO',
                'convergence_iteration': converged_at_iteration,
                'early_stopped': early_stop,
                'reset_count': pso.reset_count
            })

            run_metrics.record_pso_result(
                fitness=pso.global_best_fitness,
                is_perfect_match=is_perfect,
                convergence_iter=converged_at_iteration if early_stop else None,
                path_id=i + 1,
                method='PSO',
                reset_count=pso.reset_count,
                execution_time=path_execution_time
            )

            status = "(PSO)" if is_perfect else f"({pso.global_best_fitness:.3f})"

        print(f"  Path {i + 1}: {status} |  {path_execution_time:.2f}s")

    run_metrics.end_pso_phase()

    success_count = sum(1 for r in pso_results if r['perfect_match'])
    success_rate = (success_count / len(targetPaths)) * 100
    pso_time = run_metrics.pso_end_time - run_metrics.pso_start_time

    print(f"\nRun {run_num} runcompleted:  {success_rate:.1f}% ({success_count}/{len(targetPaths)}) | "
          f"PSO {pso_time:.2f} seconds")

    return pso_results, run_metrics


def run_multiple_dqn_pso_experiments(num_runs):
    """DQN-PSO"""
    print("\n" + "=" * 100)
    print(f"DQN-PSO(standard PSO) - {num_runs}")
    print("=" * 100)
    print(f"target paths: Path 1  Path {len(targetPaths)} ({len(targetPaths)})")
    print(f"Run: {num_runs}")
    print(f": Light: {BOUNDS['light']}, Temp: {BOUNDS['temp']}, Moisture: {BOUNDS['moisture']}")
    print("=" * 100)

    similar_group, isolated_group = group_paths_by_similarity(targetPaths)

    similar_paths_display = [idx + 1 for idx in similar_group]
    isolated_paths_display = [idx + 1 for idx in isolated_group]

    print(f"Similar path group: {similar_paths_display}")
    print(f"Isolated path group: {isolated_paths_display}\n")

    all_run_results = []
    all_run_metrics = []
    total_experiment_start = time.time()

    for run_num in range(1, num_runs + 1):
        pso_results, run_metrics = run_single_experiment(run_num, similar_group, isolated_group)
        all_run_results.append(pso_results)
        all_run_metrics.append(run_metrics)

    total_experiment_time = time.time() - total_experiment_start

    print(f"\n{'=' * 100}")
    print(f"All {num_runs} runcompleted!")
    print(f": {total_experiment_time:.2f} seconds ({total_experiment_time / 60:.2f} minutes)")
    print(f"{'=' * 100}\n")

    return all_run_results, all_run_metrics


if __name__ == "__main__":
    print("=" * 100)
    print("DQN-PSO(standard PSO) - ")
    print("=" * 100)
    print(f"Current configuration: Run = {NUM_RUNS}")
    print(f"Path : Path 1 - Path {len(targetPaths)}")
    print("=" * 100)

    # Run
    if len(sys.argv) > 1:
        try:
            NUM_RUNS = int(sys.argv[1])
            print(f"Read from command line: Run {NUM_RUNS}")
        except ValueError:
            print(f"Invalid command-line argument, using default number of runs {NUM_RUNS}")

    print("\n...")

    all_run_results, all_run_metrics = run_multiple_dqn_pso_experiments(num_runs=NUM_RUNS)

    print("\nExcel...")
    excel_filename = export_multiple_runs_to_excel(all_run_results, all_run_metrics, NUM_RUNS)

    print(f"\nProgram completed!")
    print(f"Excel: {excel_filename}")
    print("=" * 100)