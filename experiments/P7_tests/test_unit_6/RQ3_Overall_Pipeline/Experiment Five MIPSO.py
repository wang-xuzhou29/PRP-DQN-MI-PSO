import os
import random
import numpy as np
import time
import psutil
import math
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

NUM_RUNS = 20  # 默认实验运行次数

# === 新的三维范围设置 ===
LIGHT_MIN = 1
LIGHT_MAX = 100
MOISTURE_MIN = 1
MOISTURE_MAX = 100
TEMP_MIN = 1
TEMP_MAX = 60

BOUNDS = {
    'light': (LIGHT_MIN, LIGHT_MAX),
    'temp': (TEMP_MIN, TEMP_MAX),
    'moisture': (MOISTURE_MIN, MOISTURE_MAX)
}


def clip_state(state):
    """将状态限制在各维度的边界内"""
    return np.array([
        np.clip(state[0], BOUNDS['light'][0], BOUNDS['light'][1]),
        np.clip(state[1], BOUNDS['temp'][0], BOUNDS['temp'][1]),
        np.clip(state[2], BOUNDS['moisture'][0], BOUNDS['moisture'][1])
    ])


def execute_Tr(position):
    """执行目标函数并返回触发的路径"""
    x = int(np.clip(position[0], BOUNDS['light'][0], BOUNDS['light'][1]))
    temp = int(np.clip(position[1], BOUNDS['temp'][0], BOUNDS['temp'][1]))
    z = int(np.clip(position[2], BOUNDS['moisture'][0], BOUNDS['moisture'][1]))
    return category1_multivariable_control(x, temp, z)


# === 目标函数 ===
def category1_multivariable_control(x, y, z):
    # 确保数值在范围内
    x = max(STATE_MIN_X, min(STATE_MAX_X, x))
    y = max(STATE_MIN_Y, min(STATE_MAX_Y, y))
    z = max(STATE_MIN_Z, min(STATE_MAX_Z, z))

    triggered = set()

    if (25 < x < 75) != (25 < x < 80):
        triggered.add(1)
    if (25 < x < 75) != (30 < x < 75):
        triggered.add(2)

    if (25 < y < 75) != (40 < y < 75):
        triggered.add(3)
    if (25 < y < 75) != (50 < y < 75):
        triggered.add(4)

    if (20 < z < 50) != (30 < z < 50):
        triggered.add(5)
    if (20 < z < 50) != (20 < z < 30):
        triggered.add(6)

    if (x > 40 and x < 60 and y > 35 and y < 65) != (x > 40 and x < 75 and y > 35 and y < 65):
        triggered.add(7)
    if (x > 40 and x < 60 and y > 35 and y < 65) != (x > 40 and x < 60 and y > 20 and y < 65):
        triggered.add(8)

    if (x > 45 and x < 55 and z > 25 and z < 35) != (x > 45 and x < 55 and z > 10 and z < 35):
        triggered.add(9)

    if (y > 40 and y < 60 and z > 25 and z < 35) != (y > 40 and y < 85 and z > 25 and z < 35):
        triggered.add(10)
    if (y > 40 and y < 60 and z > 25 and z < 35) != (y > 40 and y < 60 and z > 30 and z < 40):
        triggered.add(11)

    if (x > 30 and x < 52) != (x > 30 and x < 70):
        triggered.add(12)
    if (x > 30 and x < 52) != (x > 30 and x < 80):
        triggered.add(13)

    if (y > 30 and y < 52) != (y > 30 and y < 70):
        triggered.add(14)
    if (y > 30 and y < 52) != (y > 30 and y < 80):
        triggered.add(15)

    if (z > 20 and z < 32) != (z > 20 and z < 40):
        triggered.add(16)
    if (z > 20 and z < 32) != (z > 20 and z < 50):
        triggered.add(17)

    if (abs(x - 50) < 5) != (abs(x - 50) < 10):
        triggered.add(18)
    if (abs(x - 50) < 5) != (abs(x - 30) < 5):
        triggered.add(19)

    if (abs(y - 50) < 8) != (abs(y - 50) < 15):
        triggered.add(20)
    if (abs(y - 50) < 8) != (abs(y - 20) < 8):
        triggered.add(21)

    if (abs(z - 30) < 3) != (abs(z - 30) < 6):
        triggered.add(22)
    if (abs(z - 30) < 3) != (abs(z - 20) < 3):
        triggered.add(23)

    if (x > 35 and y > 35 and z > 20) != (x > 35 and y > 60 and z > 20):
        triggered.add(24)

    if (x < 75 and y < 75 and z < 45) != (x < 75 and y < 40 and z < 45):
        triggered.add(25)
    if (x < 75 and y < 75 and z < 45) != (x < 75 and y < 30 and z < 45):
        triggered.add(26)

    if (x + y + z > 20 and x + y + z < 180) != (x + y + z > 50 and x + y + z < 180):
        triggered.add(27)
    if (x + y + z > 20 and x + y + z < 180) != (x + y + z > 60 and x + y + z < 180):
        triggered.add(28)

    if (x * y > 1000 and x * y < 4000) != (x * y > 1500 and x * y < 4000):
        triggered.add(29)
    if (x * y > 1000 and x * y < 4000) != (x * y > 1000 and x * y < 5000):
        triggered.add(30)

    if (x * z > 1200 and x * z < 3400) != (x * z > 1500 and x * z < 3400):
        triggered.add(31)
    if (x * z > 1200 and x * z < 3400) != (x * z > 1200 and x * z < 3000):
        triggered.add(32)

    if (y * z > 1000 and y * z < 3000) != (y * z > 1500 and y * z < 3000):
        triggered.add(33)
    if (y * z > 1000 and y * z < 3000) != (y * z > 1000 and y * z < 3500):
        triggered.add(34)

    if ((x + y + z) / 3 > 35 and (x + y + z) / 3 < 55) != ((x + y + z) / 3 > 40 and (x + y + z) / 3 < 55):
        triggered.add(35)
    if ((x + y + z) / 3 > 35 and (x + y + z) / 3 < 55) != ((x + y + z) / 3 > 35 and (x + y + z) / 3 < 40):
        triggered.add(36)

    if (math.sqrt(x ** 2 + y ** 2 + z ** 2) > 180) != (math.sqrt(x ** 2 + y ** 2 + z ** 2) > 100):
        triggered.add(37)
    if (math.sqrt(x ** 2 + y ** 2 + z ** 2) > 180) != (math.sqrt(x ** 2 + y ** 2 + z ** 2) > 120):
        triggered.add(38)

    if (x / y > 1 and x / y < 3) != (x / y > 1 and x / y < 1.2):
        triggered.add(39)
    if (x / y > 1 and x / y < 3) != (x / y > 1 and x / y < 2):
        triggered.add(40)

    if (x / z > 1.5 and x / z < 4.5) != (x / z > 1.8 and x / z < 4.5):
        triggered.add(41)
    if (x / z > 1.5 and x / z < 4.5) != (x / z > 1.5 and x / z < 3):
        triggered.add(42)

    if (y / z > 1.2 and y / z < 4.0) != (y / z > 1.4 and y / z < 4.0):
        triggered.add(43)
    if (y / z > 1.2 and y / z < 4.0) != (y / z > 1.2 and y / z < 3.5):
        triggered.add(44)

    if ((x - 50) + (y - 50) > 10 and (x - 50) - (y - 50) < 10) != (
            (x - 50) + (y - 50) > 20 and (x - 50) - (y - 50) < 10):
        triggered.add(45)
    if ((x - 50) + (y - 50) > 10 and (x - 50) - (y - 50) < 10) != (
            (x - 50) + (y - 50) > 50 and (x - 50) - (y - 50) < 10):
        triggered.add(46)

    if ((x - 50) + (z - 30) * 2 > 15 and (x - 50) + (z - 30) * 2 < 15) != (
            (x - 50) + (z - 30) * 2 > 10 and (x - 50) + (z - 30) * 2 < 15):
        triggered.add(47)
    if ((x - 50) + (z - 30) * 2 > 15 and (x - 50) + (z - 30) * 2 < 15) != (
            (x - 50) + (z - 30) * 2 > 15 and (x - 50) + (z - 30) * 2 < 40):
        triggered.add(48)

    if ((y - 50) + (z - 30) > 20 and (y - 50) + (z - 30) < 20) != (
            (y - 50) + (z - 30) > 10 and (y - 50) + (z - 30) < 20):
        triggered.add(49)
    if ((y - 50) + (z - 30) > 20 and (y - 50) + (z - 30) < 20) != (
            (y - 50) + (z - 30) > 20 and (y - 50) + (z - 30) < 40):
        triggered.add(50)

    if (abs((x - 50) - (y - 50)) < 8) != (abs((x - 50) - (y - 50)) < 30):
        triggered.add(51)
    if (abs((x - 50) - (y - 50)) < 8) != (abs((x - 45) - (y - 50)) < 20):
        triggered.add(52)

    if (abs((x - 50) - (z - 30) * 2) < 10) != (abs((x - 50) - (z - 30) * 2) < 30):
        triggered.add(53)
    if (abs((x - 50) - (z - 30) * 2) < 10) != (abs((x - 80) - (z - 30) * 2) < 10):
        triggered.add(54)

    if (abs((y - 50) - (z - 30)) < 5) != (abs((y - 50) - (z - 30)) < 20):
        triggered.add(55)
    if (abs((y - 50) - (z - 30)) < 5) != (abs((y - 80) - (z - 30)) < 5):
        triggered.add(56)

    if (x / (y + 10) > 0.7 and x / (y + 10) < 1.1) != (x / (y + 10) > 0.8 and x / (y + 10) < 1.5):
        triggered.add(57)
    if (x / (y + 10) > 0.7 and x / (y + 10) < 1.1) != (x / (y + 10) > 0.5 and x / (y + 10) < 1.0):
        triggered.add(58)

    if (y / (z + 20) > 1.0 and y / (z + 20) < 1.8) != (y / (z + 20) > 1.5 and y / (z + 20) < 1.8):
        triggered.add(59)

    if (z / (x / 2) > 0.8 and z / (x / 2) < 1.4) != (z / (x / 2) > 1.5 and z / (x / 2) < 1.4):
        triggered.add(60)
    if (z / (x / 2) > 0.8 and z / (x / 2) < 1.4) != (z / (x / 2) > 0.8 and z / (x / 2) < 1.9):
        triggered.add(61)

    if (x * y * z > 50000 and x * y * z < 150000) != (x * y * z > 30000 and x * y * z < 150000):
        triggered.add(62)

    if (x * 0.4 + y * 0.4 + z * 0.8 > 35 and x * 0.4 + y * 0.4 + z * 0.8 < 65) != (
            x * 0.4 + y * 0.4 + z * 0.8 > 35 and x * 0.4 + y * 0.4 + z * 0.8 < 60):
        triggered.add(63)

    if ((x * y * z) ** (1 / 3) > 25) != ((x * y * z) ** (1 / 4) > 25):
        triggered.add(64)
    if ((x * y * z) ** (1 / 3) > 25) != ((x * y * z) ** (1 / 2) > 25):
        triggered.add(65)

    if ((x - 50) * (y - 50) > 200 and (x - 50) * (y - 50) < 200) != (
            (x - 50) * (y - 50) > 150 and (x - 50) * (y - 50) < 200):
        triggered.add(66)

    if ((x - 50) * (z - 30) > 150 and (x - 50) * (z - 30) < 150) != (
            (x - 50) * (z - 30) > 100 and (x - 50) * (z - 30) < 150):
        triggered.add(67)
    if ((x - 50) * (z - 30) > 150 and (x - 50) * (z - 30) < 150) != (
            (x - 30) * (z - 30) > 150 and (x - 50) * (z - 30) < 150):
        triggered.add(68)

    if (abs(x / 50 + y / 50 + z / 30 - 4) < 0.5) != (abs(x / 50 + y / 50 + z / 30 - 4) < 1):
        triggered.add(69)
    if (abs(x / 50 + y / 50 + z / 30 - 4) < 0.5) != (abs(x / 40 + y / 50 + z / 30 - 4) < 0.5):
        triggered.add(70)

    if (abs((x / 50) * (y / 50) * (z / 30) - 1) < 0.3) != (abs((x / 50) * (y / 50) * (z / 30) - 1) < 0.8):
        triggered.add(71)
    if (abs((x / 50) * (y / 50) * (z / 30) - 1) < 0.3) != (abs((x / 80) * (y / 50) * (z / 30) - 1) < 0.3):
        triggered.add(72)

    if ((x + y) / 2 > 45 and (x + y) / 2 < 55 and z > 28 and z < 69) != (
            (x + y) / 2 > 55 and (x + y) / 2 < 55 and z > 38 and z < 69):
        triggered.add(73)
    if ((x + y) / 2 > 45 and (x + y) / 2 < 55 and z > 28 and z < 69) != (
            (x + y) / 2 > 55 and (x + y) / 2 < 53 and z > 48 and z < 69):
        triggered.add(74)

    if (abs(x - y) < 10 and z > 20 and z < 60) != (abs(x - y) < 20 and z > 20 and z < 60):
        triggered.add(75)
    if (abs(x - y) < 10 and z > 20 and z < 60) != (abs(x - y) < 10 and z > 40 and z < 60):
        triggered.add(76)

    if (math.sqrt((x - 50) ** 2 + (y - 50) ** 2) < 8 and abs(z - 30) < 3) != (
            math.sqrt((x - 50) ** 2 + (y - 50) ** 2) < 10 and abs(z - 30) < 3):
        triggered.add(77)
    if (math.sqrt((x - 50) ** 2 + (y - 50) ** 2) < 8 and abs(z - 30) < 3) != (
            math.sqrt((x - 50) ** 2 + (y - 50) ** 2) < 8 and abs(z - 30) < 10):
        triggered.add(78)

    if (x / 50 > 0.9 and x / 50 < 1.1 and y / 50 > 0.9 and y / 50 < 1.1) != (
            x / 50 > 1 and x / 50 < 1.1 and y / 50 > 0.9 and y / 50 < 1.1):
        triggered.add(79)
    if (x / 50 > 0.9 and x / 50 < 1.1 and y / 50 > 0.9 and y / 50 < 1.1) != (
            x / 50 > 1.2 and x / 50 < 1.1 and y / 50 > 0.9 and y / 50 < 1.1):
        triggered.add(80)

    if (z / 30 > 0.85 and z / 30 < 1) != (z / 30 > 0.85 and z / 30 < 1.3):
        triggered.add(81)

    if ((x / 50 + y / 50 + z / 30) / 3 > 0.9 and (x / 50 + y / 50 + z / 30) / 3 < 1) != (
            (x / 50 + y / 50 + z / 30) / 3 > 0.9 and (x / 50 + y / 50 + z / 30) / 3 < 1.5):
        triggered.add(82)

    if (max(abs(x / 50 - 1), abs(y / 50 - 1), abs(z / 30 - 1)) < 0.15) != (
            max(abs(x / 50 - 1), abs(y / 50 - 1), abs(z / 30 - 1)) < 0.25):
        triggered.add(83)
    if (max(abs(x / 50 - 1), abs(y / 50 - 1), abs(z / 30 - 1)) < 0.15) != (
            max(abs(x / 25 - 1), abs(y / 50 - 1), abs(z / 30 - 1)) < 0.15):
        triggered.add(84)

    if (min(x / 50, y / 50, z / 30) > 0.8) != (min(x / 50, y / 50, z / 30) > 1):
        triggered.add(85)
    if (min(x / 50, y / 50, z / 30) > 0.8) != (min(x / 50, y / 50, z / 30) > 1.8):
        triggered.add(86)

    if (max(x / 50, y / 50, z / 30) < 1.2) != (max(x / 50, y / 50, z / 30) < 1.6):
        triggered.add(87)

    if (abs(max(x, y, z * 2) - min(x, y, z * 2)) < 20) != (abs(max(x, y, z * 3) - min(x, y, z * 3)) < 20):
        triggered.add(88)

    if (x + y + z * 1.5 > 100 and x + y + z * 1.5 < 300) != (x + y + z * 1.5 > 180 and x + y + z * 1.5 < 300):
        triggered.add(89)

    if (x * y * z > 40000 and x * y * z < 120000) != (x * y * z > 45000 and x * y * z < 120000):
        triggered.add(90)
    if (x * y * z > 40000 and x * y * z < 120000) != (x * y * z > 40000 and x * y * z < 95000):
        triggered.add(91)

    if (abs((x + y + z) / 3 - 45) < 3) != (abs((x + y + z) / 3 - 45) < 10):
        triggered.add(92)
    if (abs((x + y + z) / 3 - 45) < 3) != (abs((x + y + z) / 3 - 30) < 3):
        triggered.add(93)

    if (abs(x * y / z - 75) < 5) != (abs(x * y / z - 75) < 10):
        triggered.add(94)

    if (x < 30 or y < 30 or z < 15) != (x < 20 or y < 30 or z < 15):
        triggered.add(95)
    if (x < 30 or y < 30 or z < 15) != (x < 30 or y < 20 or z < 15):
        triggered.add(96)

    if (x > 75 or y > 75 or z > 45) != (x > 60 or y > 75 or z > 45):
        triggered.add(97)
    if (x > 75 or y > 75 or z > 45) != (x > 75 or y > 60 or z > 45):
        triggered.add(98)

    if (abs(x / y - 1.0) > 0.25) != (abs(x / y - 1.0) > 0.3):
        triggered.add(99)
    if (abs(x / y - 1.0) > 0.25) != (abs(x / y - 0.95) > 0.25):
        triggered.add(100)

    if (abs(x / z - 1.8) > 0.3) != (abs(x / z - 1.8) > 0.4):
        triggered.add(101)

    if (abs(y / z - 1.7) > 0.3) != (abs(y / z - 1.6) > 0.3):
        triggered.add(102)

    if (x + y + z < 110 or x + y + z > 190) != (x + y + z < 115 or x + y + z > 190):
        triggered.add(103)

    if (x * y * z < 60000 or x * y * z > 140000) != (x * y * z < 60000 or x * y * z > 135000):
        triggered.add(104)

    if (x < 35 and y < 35) != (x < 30 and y < 35):
        triggered.add(105)
    if (x < 35 and y < 35) != (x < 35 and y < 30):
        triggered.add(106)

    if (y < 35 and z < 20) != (y < 30 and z < 20):
        triggered.add(107)

    if (x < 25 and y < 25 and z < 15) != (x < 22 and y < 25 and z < 15):
        triggered.add(108)
    if (x < 25 and y < 25 and z < 15) != (x < 25 and y < 22 and z < 15):
        triggered.add(109)

    return triggered


# 目标路径定义
targetPaths = [
    {4, 6, 11, 12, 13, 16, 17, 18, 21, 24, 25, 26, 36, 39, 41, 43, 48, 51, 52, 54, 55, 58, 60, 64, 68, 69, 70, 71, 72,
     73, 74, 75, 81, 82, 83, 85, 86, 87, 89, 91, 93, 99, 102},
    {6, 10, 12, 13, 14, 15, 16, 17, 18, 20, 22, 25, 26, 36, 41, 45, 46, 47, 49, 51, 52, 54, 55, 57, 59, 60, 64, 69, 70,
     71, 72, 76, 81, 82, 86, 87, 88, 89, 91, 92, 98, 102},
    {6, 7, 11, 12, 13, 14, 15, 16, 17, 21, 24, 25, 26, 36, 39, 41, 43, 48, 49, 51, 52, 54, 55, 58, 60, 64, 67, 68, 72,
     75, 82, 86, 87, 89, 92, 97, 100, 102, 104},
    {6, 11, 12, 13, 14, 15, 16, 17, 19, 21, 24, 25, 26, 36, 48, 49, 53, 56, 57, 61, 64, 68, 69, 70, 71, 72, 73, 74, 76,
     78, 81, 82, 83, 86, 87, 89, 91, 92, 101},
    {6, 7, 13, 14, 15, 16, 17, 20, 22, 24, 25, 26, 30, 36, 39, 48, 49, 51, 52, 53, 56, 58, 59, 60, 64, 66, 68, 72, 75,
     81, 82, 86, 87, 88, 89, 92, 97, 100, 104},
    {4, 6, 12, 13, 16, 17, 19, 21, 22, 24, 25, 26, 33, 36, 39, 41, 43, 47, 51, 52, 54, 55, 60, 64, 69, 72, 73, 74, 76,
     78, 81, 82, 84, 85, 86, 89, 93, 94, 100},
    {5, 7, 11, 12, 13, 14, 15, 21, 22, 24, 25, 26, 33, 36, 39, 47, 51, 52, 53, 54, 55, 58, 59, 64, 69, 70, 71, 72, 75,
     82, 85, 86, 87, 88, 89, 91, 92, 97, 100},
    {6, 10, 14, 15, 16, 17, 18, 20, 22, 24, 25, 26, 31, 36, 49, 51, 52, 53, 55, 58, 59, 61, 64, 69, 73, 74, 75, 81, 82,
     83, 85, 86, 87, 89, 93, 94, 99, 100},
    {5, 7, 11, 13, 14, 15, 21, 23, 24, 25, 26, 30, 36, 39, 48, 51, 53, 54, 55, 58, 59, 60, 64, 66, 71, 72, 75, 82, 85,
     86, 87, 88, 89, 91, 92, 97, 99, 102},
    {6, 10, 12, 13, 15, 16, 17, 19, 23, 25, 26, 36, 41, 46, 50, 51, 52, 54, 55, 58, 59, 60, 64, 71, 72, 75, 81, 82, 86,
     87, 88, 89, 91, 92, 98, 99, 100},
    {4, 6, 11, 12, 13, 16, 17, 19, 21, 22, 24, 25, 26, 36, 43, 47, 54, 55, 60, 64, 69, 72, 73, 74, 76, 78, 80, 81, 82,
     83, 85, 86, 89, 93, 94, 101, 102},
    {5, 6, 11, 14, 15, 19, 20, 23, 24, 25, 26, 31, 36, 41, 51, 52, 54, 55, 57, 59, 60, 64, 69, 72, 73, 74, 75, 77, 81,
     82, 83, 85, 86, 88, 89, 93, 102},
    {5, 6, 11, 14, 15, 19, 21, 23, 24, 25, 26, 31, 36, 41, 54, 56, 57, 59, 60, 63, 64, 69, 72, 73, 74, 76, 79, 80, 81,
     82, 84, 85, 86, 88, 89, 93, 94},
    {5, 6, 7, 10, 13, 14, 15, 20, 23, 25, 26, 30, 37, 46, 48, 49, 51, 52, 53, 54, 55, 59, 60, 64, 72, 76, 81, 82, 85,
     86, 87, 88, 89, 97, 98, 104},
    {5, 9, 12, 13, 14, 15, 19, 20, 25, 26, 31, 33, 36, 45, 46, 51, 52, 53, 55, 57, 59, 60, 63, 64, 69, 72, 75, 82, 83,
     87, 88, 89, 93, 98, 101},
    {3, 4, 6, 7, 12, 13, 16, 17, 20, 22, 24, 26, 33, 36, 39, 48, 51, 54, 55, 57, 60, 64, 68, 69, 70, 72, 73, 74, 81, 82,
     87, 89, 93, 94, 97},
    {1, 5, 10, 13, 14, 15, 20, 22, 30, 36, 37, 39, 48, 51, 52, 54, 55, 58, 59, 64, 71, 72, 75, 82, 85, 86, 87, 88, 89,
     92, 97, 98, 100},
    {5, 7, 13, 14, 15, 20, 25, 26, 30, 33, 36, 42, 47, 51, 52, 54, 55, 58, 59, 64, 69, 70, 71, 72, 75, 82, 87, 88, 89,
     91, 92, 97, 98},
    {3, 4, 6, 16, 17, 19, 20, 23, 24, 26, 33, 35, 36, 39, 41, 43, 51, 52, 54, 55, 58, 60, 63, 64, 72, 75, 81, 83, 89,
     92, 99},
    {12, 13, 15, 23, 25, 26, 30, 31, 33, 36, 37, 42, 44, 46, 49, 54, 56, 64, 69, 70, 71, 72, 82, 87, 88, 89, 91, 92, 97,
     98},
    {6, 7, 13, 14, 15, 17, 20, 24, 30, 32, 37, 39, 41, 43, 50, 51, 52, 54, 55, 58, 60, 64, 70, 75, 82, 86, 87, 97, 100,
     102},
    {6, 12, 13, 15, 17, 18, 30, 34, 37, 46, 48, 50, 51, 52, 53, 55, 58, 59, 61, 64, 66, 68, 82, 86, 87, 99, 100, 101,
     102},
    {3, 4, 6, 13, 16, 17, 22, 26, 33, 36, 39, 40, 48, 53, 60, 64, 68, 69, 70, 72, 73, 74, 81, 82, 87, 89, 93, 97, 101},
    {3, 4, 6, 8, 12, 13, 16, 17, 19, 21, 22, 29, 35, 39, 40, 47, 51, 54, 57, 60, 62, 64, 71, 81, 89, 92, 96, 101, 103},
    {2, 6, 10, 14, 15, 16, 17, 19, 22, 25, 26, 36, 50, 53, 55, 59, 63, 64, 72, 73, 74, 81, 82, 87, 89, 93, 95, 98},
    {3, 4, 5, 9, 12, 13, 19, 20, 23, 24, 26, 35, 39, 51, 52, 53, 56, 58, 60, 62, 64, 71, 75, 88, 89, 90, 92, 103},
    {6, 15, 16, 17, 37, 38, 39, 50, 51, 53, 54, 55, 58, 59, 60, 64, 69, 75, 81, 86, 88, 100, 102},
    {3, 4, 31, 35, 48, 49, 51, 52, 58, 62, 64, 69, 71, 82, 89, 90, 92, 95, 99, 100, 103, 106},
    {3, 4, 12, 13, 23, 26, 35, 39, 42, 51, 54, 55, 57, 62, 64, 71, 87, 89, 92, 103, 107},
    {3, 4, 6, 17, 19, 26, 29, 31, 33, 35, 47, 62, 63, 64, 71, 87, 89, 92, 103, 105, 106},
    {7, 12, 13, 14, 15, 21, 25, 26, 35, 39, 51, 52, 56, 58, 65, 87, 89, 92, 97, 100},
    {3, 4, 26, 28, 51, 52, 53, 54, 55, 58, 59, 65, 99, 100, 106, 107},
    {3, 4, 26, 27, 28, 41, 51, 52, 54, 55, 59, 60, 65, 106, 107},
    {27, 28, 39, 40, 42, 51, 52, 53, 54, 55, 57, 65, 108},
    {21, 27, 28, 53, 54, 56, 58, 59, 65, 88, 109},
]


# === 增强版指标收集器 ===
class MetricsCollector:
    def __init__(self):
        self.pso_start_time = None
        self.pso_end_time = None
        self.perfect_solutions_count = 0
        self.total_paths_count = 0
        self.final_fitness_scores = []
        self.pso_convergence_iterations = []
        self.pso_reset_counts = []
        self.path_execution_times = []

    def start_pso_phase(self):
        self.pso_start_time = time.time()

    def end_pso_phase(self):
        self.pso_end_time = time.time()

    def record_pso_result(self, fitness, is_perfect_match, convergence_iter=None, path_id=None, method='MI-PSO',
                          reset_count=0, execution_time=0):
        self.final_fitness_scores.append(fitness)
        self.total_paths_count += 1
        self.pso_reset_counts.append(reset_count)
        self.path_execution_times.append(execution_time)

        if is_perfect_match:
            self.perfect_solutions_count += 1
        if convergence_iter is not None:
            self.pso_convergence_iterations.append(convergence_iter)


# === 粒子类 ===
class Particle:
    def __init__(self):
        self.position = np.array([
            np.random.uniform(BOUNDS['light'][0], BOUNDS['light'][1]),
            np.random.uniform(BOUNDS['temp'][0], BOUNDS['temp'][1]),
            np.random.uniform(BOUNDS['moisture'][0], BOUNDS['moisture'][1])
        ])
        self.velocity = np.array([
            np.random.uniform(-5, 5),
            np.random.uniform(-3, 3),
            np.random.uniform(-5, 5)
        ])
        self.best_position = self.position.copy()
        self.best_fitness = 0
        self.fitness = 0


# === 单独的 MI-PSO 优化器类（已完全修复随机性重评估Bug） ===
class MIPSO:
    def __init__(self, target_path, swarm_size=20, max_iterations=3000, g2=10, th_cv=1.2):
        self.target_path = target_path
        self.swarm_size = swarm_size
        self.max_iterations = max_iterations
        self.g2 = g2
        self.th_cv = th_cv

        # 完全随机初始化
        self.particles = [Particle() for _ in range(swarm_size)]
        self.global_best_position = None
        self.global_best_fitness = 0
        self.global_best_triggered = set()  # 核心修复：记录最高分对应的真实触发路径
        self.reset_count = 0

        # 初始化适应度评估
        for particle in self.particles:
            fit, trig = self.evaluate(particle.position)
            particle.fitness = fit
            particle.best_fitness = fit
            particle.best_position = particle.position.copy()

            if fit > self.global_best_fitness:
                self.global_best_fitness = fit
                self.global_best_position = particle.position.copy()
                self.global_best_triggered = trig

    def evaluate(self, position):
        """同时返回适应度和实际触发的路径"""
        try:
            triggered = execute_Tr(position)
            if self.target_path.issubset(triggered):
                return 1.0, triggered
            intersection = len(triggered & self.target_path)
            union = len(triggered | self.target_path)
            return (intersection / union if union > 0 else 0.0), triggered
        except:
            return 0.0, set()

    def calculate_cv(self, scores):
        """实现公式 25：综合变异系数"""
        N = self.swarm_size
        mean_fit = np.mean(scores)
        if mean_fit == 0: return float('inf')

        var_term = np.sum((scores - mean_fit) ** 2) / N
        adj_diff = sum(abs(scores[2 * i + 1] - scores[2 * i]) for i in range(N // 2))
        adj_term = ((2 / N) * adj_diff) ** 2

        numerator = math.sqrt(var_term + adj_term)
        denominator = 2 * np.sum(scores)

        if denominator == 0: return float('inf')
        return numerator / denominator

    def flip_mutation(self, position, rho=0.1):
        """实现公式 27：变异反转操作"""
        new_pos = np.copy(position)
        new_pos[0] = BOUNDS['light'][0] + BOUNDS['light'][1] - position[0] + rho
        new_pos[1] = BOUNDS['temp'][0] + BOUNDS['temp'][1] - position[1] + rho
        new_pos[2] = BOUNDS['moisture'][0] + BOUNDS['moisture'][1] - position[2] + rho
        return clip_state(new_pos)

    def optimize(self):
        w = 0.7
        c1 = 1.5
        c2 = 1.5

        max_velocity = np.array([
            (BOUNDS['light'][1] - BOUNDS['light'][0]) * 0.2,
            (BOUNDS['temp'][1] - BOUNDS['temp'][0]) * 0.2,
            (BOUNDS['moisture'][1] - BOUNDS['moisture'][0]) * 0.2
        ])

        c = 0
        early_stop = False
        converged_at = self.max_iterations

        for iteration in range(self.max_iterations):
            scores = np.array([p.fitness for p in self.particles])
            c += 1

            # MI-PSO 特有的停滞检测与变异 (对应 Algorithm 4 Line 20-23)
            if c == self.g2:
                cv = self.calculate_cv(scores)
                if cv <= self.th_cv:
                    self.reset_count += 1
                    l = int(round(self.swarm_size * (1 - cv)))
                    l = min(max(l, 1), self.swarm_size)

                    selected_indices = np.random.choice(self.swarm_size, l, replace=False)
                    for idx in selected_indices:
                        new_pos = self.flip_mutation(self.particles[idx].position)
                        new_fit, new_trig = self.evaluate(new_pos)

                        if new_fit > self.particles[idx].fitness:
                            self.particles[idx].position = new_pos
                            self.particles[idx].fitness = new_fit
                            if new_fit > self.particles[idx].best_fitness:
                                self.particles[idx].best_fitness = new_fit
                                self.particles[idx].best_position = new_pos.copy()
                            if new_fit > self.global_best_fitness:
                                self.global_best_fitness = new_fit
                                self.global_best_position = new_pos.copy()
                                self.global_best_triggered = new_trig  # 更新最优路径
                c = 0

                # 标准 PSO 速度与位置更新
            for particle in self.particles:
                r1, r2 = np.random.random(3), np.random.random(3)
                particle.velocity = (w * particle.velocity +
                                     c1 * r1 * (particle.best_position - particle.position) +
                                     c2 * r2 * (self.global_best_position - particle.position))
                particle.velocity = np.clip(particle.velocity, -max_velocity, max_velocity)

                particle.position += particle.velocity
                particle.position = clip_state(particle.position)

                fit, trig = self.evaluate(particle.position)
                particle.fitness = fit

                if fit > particle.best_fitness:
                    particle.best_fitness = fit
                    particle.best_position = particle.position.copy()
                if fit > self.global_best_fitness:
                    self.global_best_fitness = fit
                    self.global_best_position = particle.position.copy()
                    self.global_best_triggered = trig  # 更新最优路径

            if self.global_best_fitness >= 1.0:
                converged_at = iteration + 1
                early_stop = True
                break

        return self.global_best_position, self.global_best_fitness, self.global_best_triggered, converged_at, early_stop, self.reset_count


# === Excel导出函数 ===
def export_multiple_runs_to_excel(all_run_results, all_run_metrics, num_runs, filename=None):
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Isolated_MIPSO_{num_runs}Runs_{timestamp}.xlsx"

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    border = Border(
        left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 工作表1: 运行汇总
    ws1 = wb.active
    ws1.title = "运行汇总"
    ws1.sheet_view.showGridLines = False

    headers = ["运行次数", "成功率", "成功数量", "平均适应度", "平均迭代次数", "运行时间(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, border, center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, (results, run_metrics) in enumerate(zip(all_run_results, all_run_metrics), start=1):
        success_count = sum(1 for r in results if r['perfect_match'])
        success_rate = (success_count / len(targetPaths)) * 100
        avg_fitness = np.mean([r['fitness'] for r in results])

        iterations_list = [r['convergence_iteration'] if r.get('convergence_iteration') is not None else 10000 for r in
                           results]
        avg_iterations = np.mean(iterations_list)
        total_time = run_metrics.pso_end_time - run_metrics.pso_start_time if run_metrics.pso_end_time else 0

        row_data = [f"运行 {run_idx}", f"{success_rate:.1f}%", f"{success_count}/{len(targetPaths)}",
                    f"{avg_fitness:.4f}", f"{avg_iterations:.1f}", f"{total_time:.2f}"]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border, cell.alignment = border, center_align
            if run_idx % 2 == 0: cell.fill = alternate_fill
            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    # 工作表2: 路径统计
    ws2 = wb.create_sheet(title="路径统计")
    ws2.sheet_view.showGridLines = False
    headers2 = ["路径编号", "成功次数", "成功率", "平均适应度", "平均迭代次数", "最小迭代", "最大迭代"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, border, center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(targetPaths)):
        success_count = sum(1 for results in all_run_results if results[path_idx]['perfect_match'])
        success_rate = (success_count / num_runs) * 100
        avg_fitness = np.mean([results[path_idx]['fitness'] for results in all_run_results])

        iterations_list = [results[path_idx]['convergence_iteration'] if results[path_idx].get(
            'convergence_iteration') is not None else 10000 for results in all_run_results]
        avg_iterations = np.mean(iterations_list)

        row_data = [f"路径 {path_idx + 1}", f"{success_count}/{num_runs}", f"{success_rate:.1f}%",
                    f"{avg_fitness:.4f}", f"{avg_iterations:.1f}", np.min(iterations_list), np.max(iterations_list)]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border, cell.alignment = border, center_align
            if (path_idx + 1) % 2 == 0: cell.fill = alternate_fill
            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    # 工作表3: 最佳粒子详情
    ws3 = wb.create_sheet(title="最佳粒子详情")
    ws3.sheet_view.showGridLines = False
    headers3 = ["路径", "运行", "最佳粒子(light,temp,moisture)", "适应度", "迭代次数", "求解方法", "生成路径"]
    col_widths3 = [10, 10, 25, 12, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, border, center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(targetPaths)):
        for run_idx, results in enumerate(all_run_results, start=1):
            result = results[path_idx]
            particle_str = f"({int(result['best_position'][0])}, {int(result['best_position'][1])}, {int(result['best_position'][2])})"
            path_str = str(sorted(list(result['triggered'])))

            convergence_iter = result['convergence_iteration'] if result.get(
                'convergence_iteration') is not None else 10000

            row_data = [f"路径{path_idx + 1}", f"运行{run_idx}", particle_str, f"{result['fitness']:.4f}",
                        convergence_iter if convergence_iter < 10000 else "-", "MI-PSO", path_str]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = left_align if col == 7 else center_align

                if result['fitness'] == 1.0:
                    cell.fill = success_fill
                elif result['fitness'] < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill
            row_idx += 1

    # 保存
    output_dir = os.path.join(os.getcwd(), "results")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"\n✓ 独立 MI-PSO 消融实验结果已导出到: {filepath}")
    return filepath


def run_single_mipso_experiment(run_num):
    print(f"\n{'=' * 50}")
    print(f"开始第 {run_num} 次运行 (独立 MI-PSO)")
    print(f"{'=' * 50}")

    run_metrics = MetricsCollector()
    run_metrics.start_pso_phase()

    pso_results = []

    for i, target_path in enumerate(targetPaths):
        path_start_time = time.time()

        mipso = MIPSO(target_path, swarm_size=20, max_iterations=3000)
        # 直接接收并使用最佳寻优过程中真实触发的 best_trig，坚决不重新 evaluate
        best_pos, best_fit, best_trig, converged_at, early_stop, reset_count = mipso.optimize()

        path_execution_time = time.time() - path_start_time

        # 使用真实的最高分数判断是否完美
        is_perfect = (best_fit >= 1.0)

        pso_results.append({
            'target_path': target_path,
            'best_position': best_pos,
            'fitness': best_fit,
            'triggered': best_trig,
            'perfect_match': is_perfect,
            'method': 'MI-PSO',
            'convergence_iteration': converged_at,
            'early_stopped': early_stop,
            'reset_count': reset_count
        })

        run_metrics.record_pso_result(
            fitness=best_fit,
            is_perfect_match=is_perfect,
            convergence_iter=converged_at if early_stop else None,
            path_id=i + 1,
            method='MI-PSO',
            reset_count=reset_count,
            execution_time=path_execution_time
        )

        status = "✓完美(MI-PSO)" if is_perfect else f"○部分({best_fit:.3f})"
        print(f"  路径{i + 1}: {status} | 变异触发次数: {reset_count} | 耗时 {path_execution_time:.2f}s")

    run_metrics.end_pso_phase()

    success_count = sum(1 for r in pso_results if r['perfect_match'])
    success_rate = (success_count / len(targetPaths)) * 100
    pso_time = run_metrics.pso_end_time - run_metrics.pso_start_time

    print(f"\n第{run_num}次运行完成: 成功率 {success_rate:.1f}% | 耗时 {pso_time:.2f}秒")
    return pso_results, run_metrics


def run_multiple_mipso_experiments(num_runs):
    all_run_results = []
    all_run_metrics = []
    total_start = time.time()

    for run_num in range(1, num_runs + 1):
        results, metrics = run_single_mipso_experiment(run_num)
        all_run_results.append(results)
        all_run_metrics.append(metrics)

    total_time = time.time() - total_start
    print(f"\n{'=' * 50}\n全部{num_runs}次独立 MI-PSO 运行完成! 总耗时: {total_time:.2f}秒\n{'=' * 50}")

    return all_run_results, all_run_metrics


if __name__ == "__main__":
    if len(sys.argv) > 1:
        try:
            NUM_RUNS = int(sys.argv[1])
        except ValueError:
            pass

    print("=" * 70)
    print(" 独立 MI-PSO 消融实验 (无 DQN 介入，完美修复随机性 Bug)")
    print("=" * 70)

    all_run_results, all_run_metrics = run_multiple_mipso_experiments(NUM_RUNS)
    export_multiple_runs_to_excel(all_run_results, all_run_metrics, NUM_RUNS)