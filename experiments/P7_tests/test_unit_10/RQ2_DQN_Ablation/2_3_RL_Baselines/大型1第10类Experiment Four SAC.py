import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
from torch.distributions import Normal
import numpy as np
import random
import time
import math
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")


# === 目标路径组 ===
targetPaths = [
    {1, 2, 3, 4, 5, 6, 11, 12, 17, 21, 22, 23, 24, 29, 31, 32, 33, 34, 36, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48, 53,
     54, 55, 56, 67, 68, 69, 70, 71, 72, 77, 78, 79, 80, 95, 96, 97, 99, 100, 101, 102, 103, 105, 106, 113, 114, 117,
     118, 123, 124, 125, 126, 129, 130, 135, 136, 139, 140, 146, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 17, 22, 23, 29, 31, 32, 33, 34, 36, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48, 53, 54, 55, 56, 57,
     58, 59, 60, 67, 68, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 101, 102, 105, 106, 113, 114, 117, 118, 123, 124,
     125, 126, 127, 128, 129, 130, 135, 136, 139, 140, 146, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 11, 12, 17, 19, 21, 22, 23, 24, 29, 31, 32, 33, 34, 36, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48,
     53, 54, 55, 56, 67, 68, 69, 70, 77, 78, 79, 80, 95, 96, 97, 99, 100, 101, 102, 103, 105, 106, 113, 114, 117, 118,
     123, 124, 125, 126, 129, 130, 135, 136, 139, 140, 145, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 11, 12, 15, 16, 17, 23, 24, 29, 31, 32, 33, 34, 36, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48, 53,
     54, 55, 56, 59, 60, 67, 68, 69, 70, 77, 78, 79, 80, 95, 96, 97, 99, 100, 101, 102, 105, 106, 113, 114, 117, 118,
     123, 124, 125, 126, 129, 130, 135, 136, 139, 140, 146, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 11, 12, 17, 22, 23, 24, 29, 31, 33, 34, 36, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48, 53, 54, 55,
     56, 67, 68, 69, 70, 71, 72, 77, 78, 79, 80, 95, 96, 97, 99, 100, 101, 102, 103, 105, 106, 113, 114, 117, 118, 123,
     124, 125, 126, 129, 130, 139, 140, 143, 144, 146, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 11, 12, 17, 22, 23, 24, 29, 31, 32, 33, 34, 35, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48, 53, 54,
     55, 56, 59, 60, 67, 68, 69, 70, 77, 78, 79, 80, 95, 96, 97, 101, 102, 105, 106, 113, 114, 117, 118, 120, 123, 124,
     125, 126, 129, 130, 135, 136, 139, 140, 142, 146, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 17, 22, 23, 29, 31, 33, 34, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48, 51, 52, 53, 54, 55, 56, 69,
     70, 71, 72, 73, 74, 77, 78, 79, 80, 95, 96, 97, 99, 100, 101, 102, 103, 105, 106, 108, 113, 114, 117, 118, 123,
     124, 125, 126, 129, 130, 139, 140, 143, 144, 146, 147, 153, 156},

    {1, 2, 5, 6, 9, 10, 17, 19, 20, 22, 23, 24, 28, 29, 35, 37, 38, 41, 42, 49, 53, 54, 55, 56, 57, 58, 63, 64, 66, 68,
     71, 72, 77, 78, 79, 80, 95, 96, 101, 102, 105, 106, 109, 113, 114, 117, 118, 120, 121, 122, 123, 124, 125, 126,
     127, 128, 129, 130, 131, 132, 142, 145, 148, 153, 155, 156},

    {1, 2, 3, 4, 5, 6, 17, 19, 22, 23, 29, 33, 34, 35, 37, 38, 39, 40, 41, 42, 53, 54, 55, 56, 57, 58, 59, 60, 67, 68,
     69, 70, 73, 74, 77, 78, 79, 80, 85, 86, 95, 96, 97, 101, 102, 105, 106, 113, 114, 117, 118, 120, 123, 124, 127,
     128, 129, 130, 135, 136, 139, 140, 142, 145, 148, 153, 156},

    {1, 2, 3, 4, 17, 19, 21, 22, 29, 31, 32, 33, 34, 37, 38, 39, 40, 47, 48, 51, 52, 53, 54, 55, 56, 57, 58, 63, 64, 69,
     70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 99, 100, 101, 102, 105, 106, 113, 114, 117, 118, 123, 124, 125, 126, 127,
     128, 129, 130, 135, 136, 145, 147, 149, 150, 153, 156},

    {1, 2, 3, 4, 5, 6, 18, 19, 22, 23, 25, 26, 29, 31, 36, 37, 38, 39, 40, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 55,
     56, 57, 58, 68, 69, 70, 71, 72, 77, 78, 79, 80, 84, 97, 101, 102, 103, 105, 106, 108, 117, 118, 119, 125, 126, 127,
     128, 129, 130, 139, 140, 143, 144, 145, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 22, 23, 25, 26, 29, 31, 35, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 53, 54, 56, 57, 58,
     61, 62, 68, 69, 70, 71, 72, 77, 78, 79, 80, 97, 101, 102, 105, 106, 108, 117, 118, 119, 123, 124, 125, 126, 127,
     128, 129, 130, 139, 140, 143, 144, 146, 147, 153, 156},

    {1, 2, 3, 4, 17, 19, 22, 29, 33, 34, 37, 38, 39, 40, 41, 42, 53, 54, 55, 56, 57, 58, 59, 60, 63, 64, 69, 70, 73, 74,
     77, 78, 79, 80, 87, 88, 95, 96, 97, 101, 102, 105, 106, 113, 114, 117, 118, 120, 123, 124, 125, 126, 127, 128, 129,
     130, 135, 136, 142, 145, 148, 149, 150, 153, 156},

    {1, 2, 3, 4, 5, 6, 18, 22, 23, 24, 29, 31, 36, 37, 38, 39, 40, 43, 44, 45, 46, 47, 48, 51, 52, 53, 54, 55, 56, 57,
     58, 67, 68, 69, 70, 71, 72, 77, 78, 79, 80, 83, 84, 97, 101, 102, 103, 105, 106, 108, 117, 118, 125, 126, 127, 128,
     129, 130, 139, 140, 143, 144, 145, 147, 153, 156},

    {1, 2, 5, 6, 17, 19, 20, 22, 23, 27, 29, 35, 37, 38, 39, 40, 41, 42, 49, 50, 53, 54, 55, 56, 57, 58, 63, 64, 66, 67,
     68, 69, 70, 71, 72, 77, 78, 79, 80, 95, 96, 101, 102, 105, 106, 113, 114, 117, 118, 120, 123, 124, 125, 126, 127,
     128, 129, 130, 131, 132, 142, 145, 148, 153, 156},

    {1, 2, 3, 4, 13, 14, 17, 22, 29, 31, 32, 33, 34, 37, 38, 39, 40, 51, 52, 53, 54, 55, 56, 57, 58, 63, 64, 69, 70, 75,
     76, 77, 78, 79, 80, 89, 90, 93, 94, 95, 96, 97, 99, 100, 101, 102, 105, 106, 112, 113, 114, 117, 118, 125, 126,
     127, 128, 129, 130, 145, 147, 149, 150, 153, 156},

    {1, 2, 3, 4, 7, 8, 13, 14, 17, 21, 22, 27, 29, 33, 34, 37, 38, 39, 40, 51, 52, 56, 57, 58, 63, 64, 69, 70, 77, 78,
     79, 80, 87, 88, 89, 90, 92, 93, 94, 95, 96, 97, 99, 100, 101, 102, 105, 106, 111, 112, 113, 114, 117, 118, 125,
     126, 127, 128, 129, 130, 145, 147, 149, 150, 156},

    {1, 2, 5, 6, 9, 10, 17, 19, 20, 23, 24, 28, 30, 35, 41, 42, 49, 53, 54, 55, 56, 57, 58, 63, 64, 66, 68, 71, 72, 77,
     78, 79, 80, 95, 96, 101, 102, 105, 106, 109, 110, 113, 114, 117, 118, 121, 122, 123, 124, 125, 126, 127, 128, 129,
     130, 131, 132, 142, 145, 148, 153, 155, 156},

    {1, 2, 3, 4, 5, 6, 17, 22, 23, 24, 28, 29, 35, 37, 38, 50, 53, 54, 56, 57, 58, 63, 64, 66, 68, 71, 72, 77, 78, 79,
     80, 89, 90, 95, 96, 101, 102, 105, 106, 109, 113, 114, 117, 118, 121, 122, 123, 124, 125, 126, 127, 128, 129, 130,
     131, 132, 138, 142, 148, 151, 153, 155, 156},

    {1, 2, 3, 4, 5, 6, 11, 12, 15, 16, 17, 19, 23, 24, 28, 29, 36, 37, 38, 39, 40, 49, 53, 54, 55, 56, 63, 64, 69, 70,
     79, 80, 82, 89, 90, 95, 96, 101, 102, 103, 105, 106, 113, 114, 116, 117, 118, 121, 122, 123, 124, 125, 126, 129,
     130, 131, 132, 133, 134, 138, 145, 147, 156},

    {1, 2, 3, 4, 5, 6, 17, 19, 20, 22, 23, 27, 29, 37, 38, 39, 40, 49, 50, 51, 52, 53, 54, 55, 56, 63, 64, 67, 68, 69,
     70, 77, 78, 79, 80, 87, 88, 89, 90, 95, 96, 97, 101, 102, 103, 105, 106, 113, 114, 116, 117, 118, 123, 124, 125,
     126, 129, 130, 143, 144, 145, 147, 154, 156},

    {1, 2, 3, 4, 7, 8, 19, 20, 21, 22, 25, 26, 28, 30, 35, 39, 40, 43, 44, 46, 53, 54, 55, 56, 57, 58, 61, 62, 65, 66,
     73, 74, 77, 78, 79, 80, 101, 102, 104, 105, 106, 109, 110, 117, 118, 119, 121, 122, 123, 124, 125, 126, 127, 129,
     130, 131, 132, 145, 147, 153, 155, 156},

    {1, 2, 3, 4, 5, 6, 19, 21, 22, 28, 29, 35, 37, 38, 39, 40, 43, 44, 45, 46, 53, 54, 57, 58, 61, 62, 66, 73, 74, 77,
     78, 79, 80, 101, 102, 103, 104, 105, 106, 117, 118, 121, 122, 123, 124, 125, 126, 127, 128, 129, 130, 131, 132,
     133, 134, 137, 145, 147, 153, 155, 156},

    {1, 2, 5, 6, 17, 23, 29, 33, 34, 35, 37, 38, 41, 42, 43, 44, 53, 54, 55, 56, 57, 58, 59, 60, 67, 68, 69, 70, 73, 74,
     79, 80, 95, 96, 98, 101, 102, 105, 106, 109, 113, 114, 117, 118, 120, 123, 124, 127, 128, 129, 130, 131, 132, 135,
     136, 139, 140, 142, 146, 153, 156},

    {1, 2, 3, 4, 5, 6, 22, 25, 26, 29, 31, 32, 35, 37, 38, 41, 42, 43, 44, 45, 46, 47, 48, 50, 53, 54, 56, 57, 58, 59,
     60, 69, 70, 79, 80, 101, 102, 105, 106, 109, 117, 118, 123, 124, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136,
     139, 140, 141, 146, 147, 153, 156},

    {1, 2, 17, 19, 27, 29, 35, 37, 38, 39, 40, 41, 42, 49, 50, 53, 54, 56, 57, 58, 63, 64, 66, 67, 68, 72, 77, 78, 79,
     80, 95, 96, 101, 102, 105, 106, 107, 109, 113, 114, 117, 118, 123, 124, 125, 126, 127, 129, 130, 131, 132, 138,
     142, 145, 149, 150, 153, 155, 156},

    {1, 2, 3, 4, 5, 6, 7, 8, 13, 14, 15, 16, 23, 25, 26, 28, 29, 35, 37, 38, 39, 40, 43, 44, 45, 46, 55, 56, 57, 58, 66,
     69, 70, 77, 78, 79, 80, 81, 92, 103, 104, 105, 106, 117, 118, 119, 121, 122, 125, 126, 127, 128, 129, 130, 131,
     132, 145, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 17, 19, 20, 21, 22, 23, 24, 28, 29, 36, 37, 38, 39, 40, 49, 50, 53, 54, 67,
     68, 69, 70, 87, 88, 95, 96, 98, 101, 102, 103, 105, 106, 113, 114, 115, 116, 117, 118, 121, 122, 129, 130, 145,
     147, 156},

    {1, 2, 3, 4, 5, 6, 7, 8, 13, 14, 17, 23, 27, 29, 36, 37, 38, 39, 40, 49, 50, 63, 64, 67, 68, 69, 70, 79, 80, 84, 85,
     86, 87, 88, 89, 90, 91, 92, 95, 96, 97, 101, 102, 103, 105, 106, 113, 114, 116, 117, 118, 129, 130, 145, 147, 154,
     156},

    {3, 4, 5, 6, 22, 28, 30, 35, 45, 46, 53, 54, 55, 56, 63, 64, 65, 66, 71, 72, 73, 74, 77, 78, 79, 80, 103, 109, 110,
     117, 118, 121, 122, 123, 124, 125, 126, 127, 128, 131, 132, 137, 141, 147, 152, 153, 155, 156},

]

# === 配置（新范围：温度 1~65，电压 100~205，流量 1~20） ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'ACTION_DIM': 3,
    'MIN_VALUES': np.array([1.0, 100.0, 1.0], dtype=np.float32),
    'MAX_VALUES': np.array([65.0, 205.0, 20.0], dtype=np.float32),
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 3,
    'REPLAY_BATCH_SIZE': 64,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'ACTOR_LR': 3e-4,
    'CRITIC_LR': 3e-4,
    'ALPHA_LR': 3e-4,
    'GAMMA': 0.99,
    'TAU': 0.005,
    'NUM_RUNS': 20,
    'TOP_K_SAMPLES': 20,
    'TARGET_PATHS': targetPaths
}

# ==================== 第10类规则函数（完整） ====================
def section10_comprehensive_hybrid_control(x, y, z):
    """第10类: 综合混合控制 (温度x, 电压y, 流量z) - 158个分支"""
    triggered = set()

    # -------------------------- 1-40：综合系统基础检查 --------------------------
    # 1-2：温度基础参数正常
    if [(30 < x < 70)] != [(30 < x * 78 < 70)]:
        triggered.add(1)
    if [(30 < x < 70)] != [(30 < 122 < 70)]:
        triggered.add(2)

    # 3-4：电压基础参数正常
    if [(140 < y < 210)] != [(140 < 678 < 210)]:
        triggered.add(3)
    if [(140 < y < 210)] != [(140 < y * 88 < 210)]:
        triggered.add(4)

    # 5-6：流量基础参数正常
    if [(8 < z < 16)] != [(8 < z * 67 < 16)]:
        triggered.add(5)
    if [(8 < z < 16)] != [(8 < 566 < 16)]:
        triggered.add(6)

    # 7-8：温度电压核心区间协调（45-55 & 160-190）
    if [(45 < x < 55 and 160 < y < 190)] != [(45 < x < 55 and 160 < y * 8 < 190)]:
        triggered.add(7)
    if [(45 < x < 55 and 160 < y < 190)] != [(45 < x * 6 < 55 and 160 < y < 190)]:
        triggered.add(8)

    # 9-10：温度流量匹配良好（42-58 & 11-14）
    if [(42 < x < 58 and 11 < z < 14)] != [(42 < x < 58 and 11 < z * 78 < 14)]:
        triggered.add(9)
    if [(42 < x < 58 and 11 < z < 14)] != [(42 < x < 58 and 11 < z * 8 < 14)]:
        triggered.add(10)

    # 11-12：电压流量协调稳定（155-195 & 11-14）
    if [(155 < y < 195 and 11 < z < 14)] != [(155 < y < 195 and 11 < z * 78 < 14)]:
        triggered.add(11)
    if [(155 < y < 195 and 11 < z < 14)] != [(155 < y * 66 < 195 and 11 < z < 14)]:
        triggered.add(12)

    # 13-14：温度精确控制（48-52）
    if [(48 < x < 52)] != [(48 < x * 78 < 52)]:
        triggered.add(13)
    if [(48 < x < 52)] != [(48 < 88 < 52)]:
        triggered.add(14)

    # 15-16：电压精确控制（173-177）
    if [(173 < y < 177)] != [(173 < y * 7 < 177)]:
        triggered.add(15)
    if [(173 < y < 177)] != [(173 < 88 < 177)]:
        triggered.add(16)

    # 17-18：流量精确控制（12.2-12.8）
    if [(12.2 < z < 12.8)] != [(12.2 < z < 128)]:
        triggered.add(17)
    if [(12.2 < z < 12.8)] != [(10 < z < 12.8)]:
        triggered.add(18)

    # 19-20：温度稳定性良好（偏离目标50 < 2）
    if [(abs(x - 50) < 2)] != [(abs(x - 50) < 12)]:
        triggered.add(19)
    if [(abs(x - 50) < 2)] != [(abs(x - 50) < 8)]:
        triggered.add(20)

    # 21-22：电压稳定性良好（偏离目标175 < 5）
    if [(abs(y - 175) < 5)] != [(abs(y - 175) < 15)]:
        triggered.add(21)
    if [(abs(y - 175) < 5)] != [(abs(y - 175) < 51)]:
        triggered.add(22)

    # 23-24：流量稳定性良好（偏离目标12.5 < 0.3）
    if [(abs(z - 12.5) < 0.3)] != [(abs(z - 12.5) < 3)]:
        triggered.add(23)
    if [(abs(z - 12.5) < 0.3)] != [(abs(z - 12.5) < 1)]:
        triggered.add(24)

    # 25-26：所有参数在安全范围（下限：42、155、11）
    if [(x > 42 and y > 155 and z > 11)] != [(x > 42 and y > 155 and 22 > 11)]:
        triggered.add(25)
    if [(x > 42 and y > 155 and z > 11)] != [(x > 42 and y > 155 and z * 3 > 11)]:
        triggered.add(26)

    # 27-28：所有参数未超上限（上限：58、195、14）
    if [(x < 58 and y < 195 and z < 14)] != [(x < 58 and y < 195 and 10 < 14)]:
        triggered.add(27)
    if [(x < 58 and y < 195 and z < 14)] != [(x < 58 and y < 195 and z * 78 < 14)]:
        triggered.add(28)

    # 29-30：综合系统指标正常（x+y+10z：260-300）
    if [(260 < x + y + z * 10 < 300)] != [(260 < x + y + z * 10 < 3100)]:
        triggered.add(29)
    if [(260 < x + y + z * 10 < 300)] != [(300 < x + y + z * 10 < 300)]:
        triggered.add(30)

    # 31-32：温度电压乘积正常（8000-10000）
    if [(8000 < x * y < 10000)] != [(8000 < x * y < 70000)]:
        triggered.add(31)
    if [(8000 < x * y < 10000)] != [(8000 < x * y < 12000)]:
        triggered.add(32)

    # 33-34：温度流量乘积正常（600-750）
    if [(600 < x * z < 750)] != [(600 < x * z < 7510)]:
        triggered.add(33)
    if [(600 < x * z < 750)] != [(600 < x * z < 1750)]:
        triggered.add(34)

    # 35-36：电压流量乘积正常（2100-2500）
    if [(2100 < y * z < 2500)] != [(210 < y * z < 2500)]:
        triggered.add(35)
    if [(2100 < y * z < 2500)] != [(2100 < 23 * z < 2500)]:
        triggered.add(36)

    # 37-38：平均系统参数正常（(x+y+10z)/3：86-100）
    if [(86 < (x + y + z * 10) / 3 < 100)] != [(86 < (x + y + z * 10) / 3 < 1020)]:
        triggered.add(37)
    if [(86 < (x + y + z * 10) / 3 < 100)] != [(86 < (x + y + z * 10) / 3 < 200)]:
        triggered.add(38)

    # 39-40：系统向量模长正常（sqrt(x²+y²+(10z)²) > 200）
    if [(math.sqrt(x ** 2 + y ** 2 + (z * 10) ** 2) > 200)] != [(math.sqrt(x ** 2 + y ** 2 + (z * 10) ** 2) > 2100)]:
        triggered.add(39)
    if [(math.sqrt(x ** 2 + y ** 2 + (z * 10) ** 2) > 200)] != [(math.sqrt(x ** 2 + y ** 2 + (z * 10) ** 2) > 500)]:
        triggered.add(40)

    # -------------------------- 41-80：跨域参数分析 --------------------------
    # 41-42：温度电压比理想（x/(y/4)：1.0-1.3）
    if [(1.0 < x / (y / 4) < 1.3)] != [(1.0 < x / (y / 4) < 13)]:
        triggered.add(41)
    if [(1.0 < x / (y / 4) < 1.3)] != [(1.0 < x / (y / 4) < 333)]:
        triggered.add(42)

    # 43-44：温度流量比正常（x/z：3.5-4.5）
    if [(3.5 < x / z < 4.5)] != [(3.5 < x / z < 45)]:
        triggered.add(43)
    if [(3.5 < x / z < 4.5)] != [(3.5 < x / z < 24.5)]:
        triggered.add(44)

    # 45-46：电压流量比适当（y/z：12-16）
    if [(12 < y / z < 16)] != [(12 < y / z < 30)]:
        triggered.add(45)
    if [(12 < y / z < 16)] != [(12 < y / z < 76)]:
        triggered.add(46)

    # 47-48：温度电压偏差和正常（(x-40)+(y-150)/5：15-25）
    if [(15 < (x - 40) + (y - 150) / 5 < 25)] != [(15 < (x - 40) + (y - 150) / 5 < 215)]:
        triggered.add(47)
    if [(15 < (x - 40) + (y - 150) / 5 < 25)] != [(15 < (x - 40) + (y - 150) / 5 < 125)]:
        triggered.add(48)

    # 49-50：温度流量偏差和正常（(x-40)+3(z-10)：15-25）
    if [(15 < (x - 40) + (z - 10) * 3 < 25)] != [(15 < (x - 40) + (z - 10) * 5 < 25)]:
        triggered.add(49)
    if [(15 < (x - 40) + (z - 10) * 3 < 25)] != [(15 < (x - 40) + (z - 10) * 10 < 25)]:
        triggered.add(50)

    # 51-52：电压流量偏差和正常（(y-150)/5+(z-10)：8-12）
    if [(8 < (y - 150) / 5 + (z - 10) < 12)] != [(8 < (y - 150) / 5 + (z - 10) < 121)]:
        triggered.add(51)
    if [(8 < (y - 150) / 5 + (z - 10) < 12)] != [(8 < (y - 150) / 5 + (z - 10) < 112)]:
        triggered.add(52)

    # 53-54：温度电压偏差关系平衡（|(x-50)-(y-175)/4| < 5）
    if [(abs((x - 50) - (y - 175) / 4) < 5)] != [(abs((x - 50) - (y - 175) / 4) < 115)]:
        triggered.add(53)
    if [(abs((x - 50) - (y - 175) / 4) < 5)] != [(abs((x - 50) - (y - 175) / 4) < 51)]:
        triggered.add(54)

    # 55-56：温度流量偏差关系平衡（|(x-50)-3(z-12.5)| < 6）
    if [(abs((x - 50) - (z - 12.5) * 3) < 6)] != [(abs((x - 50) - (z - 12.5) * 3) < 16)]:
        triggered.add(55)
    if [(abs((x - 50) - (z - 12.5) * 3) < 6)] != [(abs((x - 50) - (z - 12.5) * 3) < 26)]:
        triggered.add(56)

    # 57-58：电压流量偏差关系平衡（|(y-175)/14-(z-12.5)| < 2）
    if [(abs((y - 175) / 14 - (z - 12.5)) < 2)] != [(abs((y - 175) / 14 - (z - 12.5)) < 21)]:
        triggered.add(57)
    if [(abs((y - 175) / 14 - (z - 12.5)) < 2)] != [(abs((y - 175) / 14 - (z - 12.5)) < 112)]:
        triggered.add(58)

    # 59-60：调整温度电压比正常（x/(y/4+10)：0.9-1.2）
    if [(0.9 < x / (y / 4 + 10) < 1.2)] != [(0.9 < x / (y / 4 + 10) < 12)]:
        triggered.add(59)
    if [(0.9 < x / (y / 4 + 10) < 1.2)] != [(0.9 < x / (y / 4 + 10) < 112)]:
        triggered.add(60)

    # 61-62：调整电压流量比正常（y/(z+5)：9-13）
    if [(9 < y / (z + 5) < 13)] != [(9 < y / (z + 5) < 53)]:
        triggered.add(61)
    if [(9 < y / (z + 5) < 13)] != [(9 < y / (z + 5) < 131)]:
        triggered.add(62)

    # 63-64：调整流量温度比正常（z/(x/10)：2.2-2.8）
    if [(2.2 < z / (x / 10) < 2.8)] != [(2.2 < z / (x / 10) < 28)]:
        triggered.add(63)
    if [(2.2 < z / (x / 10) < 2.8)] != [(2.2 < z / (x / 10) < 128)]:
        triggered.add(64)

    # 65-66：三元系统积正常（x*y*z：20000-30000）
    if [(20000 < (x * y * z) < 30000)] != [(20000 < (x * y * z) < 56000)]:
        triggered.add(65)
    if [(20000 < (x * y * z) < 30000)] != [(20000 < (x * y * z) < 88000)]:
        triggered.add(66)

    # 67-68：加权系统和正常（0.8x+0.1y+8z：160-180）
    if [(160 < x * 0.8 + y * 0.1 + z * 8 < 180)] != [(160 < x * 0.8 + y * 1 + z * 8 < 180)]:
        triggered.add(67)
    if [(160 < x * 0.8 + y * 0.1 + z * 8 < 180)] != [(160 < x * 0.8 + y * 0.1 + z * 10 < 180)]:
        triggered.add(68)

    # 69-70：加权几何平均正常（(x/50)^0.3*(y/175)^0.4*(z/12.5)^0.3 > 0.9）
    if [((x / 50) ** 0.3 * (y / 175) ** 0.4 * (z / 12.5) ** 0.3 > 0.9)] != [((x / 50) ** 0.3 * (y / 175) ** 0.4 * (z / 12.5) ** 0.3 > 9)]:
        triggered.add(69)
    if [((x / 50) ** 0.3 * (y / 175) ** 0.4 * (z / 12.5) ** 0.3 > 0.9)] != [((x / 50) ** 0.3 * (y / 175) ** 0.4 * (z / 12.5) ** 0.3 > 19)]:
        triggered.add(70)

    # 71-72：温度电压偏差积平衡（(x-50)(y-175)/4：-40-40）
    if [((x - 50) * (y - 175) / 4 > -40 and (x - 50) * (y - 175) / 4 < 40)] != [((x - 50) * (y - 175) / 4 > -40 and (x - 50) * (y - 175) / 4 < 140)]:
        triggered.add(71)
    if [((x - 50) * (y - 175) / 4 > -40 and (x - 50) * (y - 175) / 4 < 40)] != [((x - 50) * (y - 175) / 4 > -40 and (x - 50) * (y - 175) / 4 < 640)]:
        triggered.add(72)

    # 73-74：温度流量偏差积平衡（(x-50)(z-12.5)：-15-15）
    if [((x - 50) * (z - 12.5) > -15 and (x - 50) * (z - 12.5) < 15)] != [((x - 50) * (z - 12.5) > -15 and (x - 50) * (z - 12.5) < 115)]:
        triggered.add(73)
    if [((x - 50) * (z - 12.5) > -15 and (x - 50) * (z - 12.5) < 15)] != [((x - 50) * (z - 12.5) > -15 and (x - 50) * (z - 12.5) < 151)]:
        triggered.add(74)

    # 75-76：电压流量偏差积平衡（(y-175)(z-12.5)/14：-8-8）
    if [((y - 175) / 14 * (z - 12.5) > -8 and (y - 175) / 14 * (z - 12.5) < 8)] != [((y - 175) / 14 * (z - 12.5) > -8 and (y - 175) / 14 * (z - 12.5) < 118)]:
        triggered.add(75)
    if [((y - 175) / 14 * (z - 12.5) > -8 and (y - 175) / 14 * (z - 12.5) < 8)] != [((y - 175) / 14 * (z - 12.5) > -8 and (y - 175) / 14 * (z - 12.5) < 138)]:
        triggered.add(76)

    # 77-78：归一化和接近理想（|x/50+y/175+z/12.5-3| < 0.15）
    if [(abs(x / 50 + y / 175 + z / 12.5 - 3) < 0.15)] != [(abs(x / 50 + y / 175 + z / 12.5 - 3) < 15)]:
        triggered.add(77)
    if [(abs(x / 50 + y / 175 + z / 12.5 - 3) < 0.15)] != [(abs(x / 50 + y / 175 + z / 12.5 - 3) < 325)]:
        triggered.add(78)

    # 79-80：归一化积接近理想（|(x/50)(y/175)(z/12.5)-1| < 0.1）
    if [(abs((x / 50) * (y / 175) * (z / 12.5) - 1) < 0.1)] != [(abs((x / 50) * (y / 175) * (z / 12.5) - 1) < 1)]:
        triggered.add(79)
    if [(abs((x / 50) * (y / 175) * (z / 12.5) - 1) < 0.1)] != [(abs((x / 50) * (y / 175) * (z / 12.5) - 1) < 12)]:
        triggered.add(80)

    # -------------------------- 81-158：全系统优化协调+智能预测控制 --------------------------
    # 81-82：全参数系统优化（温度49-51、电压173-177、流量12.3-12.7）
    if [(49 < x < 51 and 173 < y < 177 and 12 < z < 17)] != [(49 < x < 51 and 173 < y < 177 and 3 < z < 17)]:
        triggered.add(81)
    if [(49 < x < 51 and 173 < y < 177 and 12 < z < 17)] != [(25 < x < 51 and 173 < y < 177 and 12 < z < 17)]:
        triggered.add(82)

    # 83-84：温度电压比与流量协调（|x/(y/4)-1.14|<0.1 & |z-12.5|<0.2）
    if [(abs(x / (y / 4) - 1.14) < 0.1 and abs(z - 12.5) < 0.2)] != [(abs(x / (y / 4) - 1.14) < 0.1 and abs(z - 12.5) < 1)]:
        triggered.add(83)
    if [(abs(x / (y / 4) - 1.14) < 0.1 and abs(z - 12.5) < 0.2)] != [(abs(x / (y / 4) - 1.14) < 0.1 and abs(z - 12.5) < 2)]:
        triggered.add(84)

    # 85-86：温度电压积与流量协调（x*y：8700-8800 & 流量12.4-12.6）
    if [(8700 < x * y < 8800 and 12.4 < z < 12.6)] != [(8700 < x * y < 8800 and 12.4 < z < 126)]:
        triggered.add(85)
    if [(8700 < x * y < 8800 and 12.4 < z < 12.6)] != [(8700 < x * y < 8800 and 12.4 < z < 32.6)]:
        triggered.add(86)

    # 87-88：温度电压平均与流量协调（(x+y/4)/2：46-48 & 流量12.3-12.7）
    if [(46 < (x + y / 4) / 2 < 48 and 12.3 < z < 12.7)] != [(46 < (x + y / 4) / 2 < 48 and 12.3 < z < 127)]:
        triggered.add(87)
    if [(46 < (x + y / 4) / 2 < 48 and 12.3 < z < 12.7)] != [(46 < (x + y / 4) / 2 < 48 and 12.3 < z < 72.7)]:
        triggered.add(88)

    # 89-90：温度电压差与流量协调（|x-y/4|<6 & 流量12.3-12.7）
    if [(abs(x - y / 4) < 6 and 12.3 < z < 12.7)] != [(abs(x - y / 4) < 6 and 12.3 < z < 127)]:
        triggered.add(89)
    if [(abs(x - y / 4) < 6 and 12.3 < z < 12.7)] != [(abs(x - y / 4) < 6 and 12.3 < z < 327)]:
        triggered.add(90)

    # 91-92：温度电压距离与流量优秀（sqrt((x-50)²+(y/4-43.75)²)<2 & |z-12.5|<0.2）
    if [(math.sqrt((x - 50) ** 2 + (y / 4 - 43.75) ** 2) < 2 and abs(z - 12.5) < 0.2)] != [(math.sqrt((x - 50) ** 2 + (y / 4 - 43.75) ** 2) < 2 and abs(z - 12.5) < 2)]:
        triggered.add(91)
    if [(math.sqrt((x - 50) ** 2 + (y / 4 - 43.75) ** 2) < 2 and abs(z - 12.5) < 0.2)] != [(math.sqrt((x - 50) ** 2 + (y / 4 - 43.75) ** 2) < 2 and abs(z - 12.5) < 22)]:
        triggered.add(92)

    # 93-94：温度电压相对值协调（相对偏差<2%）
    if [(0.98 < x / 50 < 1.02 and 0.985 < y / 175 < 1.015)] != [(0.98 < x / 50 < 1.02 and 0.985 < y / 175 < 12)]:
        triggered.add(93)
    if [(0.98 < x / 50 < 1.02 and 0.985 < y / 175 < 1.015)] != [(0.98 < x / 50 < 1.02 and 0.985 < y / 175 < 10)]:
        triggered.add(94)

    # 95-96：流量相对值协调（相对偏差<1.6%）
    if [(0.984 < z / 12.5 < 1.016)] != [(0.984 < z / 12.5 < 10.16)]:
        triggered.add(95)
    if [(0.984 < z / 12.5 < 1.016)] != [(0.984 < z / 12.5 < 4.016)]:
        triggered.add(96)

    # 97-98：归一化平均协调（平均相对偏差<1.5%）
    if [(0.985 < (x / 50 + y / 175 + z / 12.5) / 3 < 1.015)] != [(0.985 < (x / 50 + y / 175 + z / 12.5) / 3 < 10.15)]:
        triggered.add(97)
    if [(0.985 < (x / 50 + y / 175 + z / 12.5) / 3 < 1.015)] != [(0.985 < (x / 50 + y / 175 + z / 12.5) / 8 < 1.015)]:
        triggered.add(98)

    # 99-100：最小相对值协调良好
    if [(min(x / 50, y / 175, z / 12.5) > 0.985)] != [(min(x / 50, y / 175, z / 12.5) > 9.85)]:
        triggered.add(99)
    if [(min(x / 50, y / 175, z / 12.5) > 0.985)] != [(min(x / 50, y / 175, z / 12.5) > 9.85)]:
        triggered.add(100)

    # 101-102：最大相对值协调良好
    if [(max(x / 50, y / 175, z / 12.5) < 1.015)] != [(max(x / 50, y / 175, z / 12.5) < 15)]:
        triggered.add(101)
    if [(max(x / 50, y / 175, z / 12.5) < 1.015)] != [(max(x / 50, y / 175, z / 12.5) < 11.5)]:
        triggered.add(102)

    # 103-104：标准化后范围协调
    if [(abs(max(x, y / 4, z * 4) - min(x, y / 4, z * 4)) < 8)] != [(abs(max(x, y / 4, z * 4) - min(x, y / 4, z * 4)) < 18)]:
        triggered.add(103)
    if [(abs(max(x, y / 4, z * 4) - min(x, y / 4, z * 4)) < 8)] != [(abs(max(x, y / 4, z * 4) - min(x, y / 4, z * 14)) < 8)]:
        triggered.add(104)

    # 105-106：综合参数协调优秀
    if [(278 < (x + y + z * 10) < 282)] != [(278 < (x + y + z * 10) < 482)]:
        triggered.add(105)
    if [(278 < (x + y + z * 10) < 282)] != [(278 < (x + y + z * 10) < 989)]:
        triggered.add(106)

    # 107-108：三元积协调优秀
    if [(24500 < (x * y * z) < 25500)] != [(24500 < (x * y * 6) < 25500)]:
        triggered.add(107)
    if [(24500 < (x * y * z) < 25500)] != [(24500 < (x * y * 2) < 25500)]:
        triggered.add(108)

    # 109-110：平均值协调优秀
    if [(abs((x + y + z * 10) / 3 - 93.3) < 1)] != [(abs((x + y + z * 10) / 3 - 93.3) < 11)]:
        triggered.add(109)
    if [(abs((x + y + z * 10) / 3 - 93.3) < 1)] != [(abs((x + y + z * 10) / 3 - 93.3) < 6)]:
        triggered.add(110)

    # 111-112：温度电压超精密协调
    if [(49.5 < x < 50.5 and 174 < y < 176)] != [(49.5 < x < 50.5 and 174 < y < 196)]:
        triggered.add(111)
    if [(49.5 < x < 50.5 and 174 < y < 176)] != [(49.5 < x < 50.5 and 174 < y < 676)]:
        triggered.add(112)

    # 113-114：流量超精密协调
    if [(12.45 < z < 12.55)] != [(12.45 < z < 128)]:
        triggered.add(113)
    if [(12.45 < z < 12.55)] != [(12.45 < z < 1255)]:
        triggered.add(114)

    # 115-116：三维距离协调完美
    if [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 16) < 16] != [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 4 * 16) < 16]:
        triggered.add(115)
    if [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 16) < 16] != [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 16) < 96]:
        triggered.add(116)

    # 117-118：连续比例协调完美
    if [(abs(x / (y / 4) / z - 0.091) < 0.003)] != [(abs(x / (y / 4) / z - 0.091) < 3)]:
        triggered.add(117)
    if [(abs(x / (y / 4) / z - 0.091) < 0.003)] != [(abs(x / (y / 4) / z - 0.091) < 1.3)]:
        triggered.add(118)

    # 119-120：检测到参数偏低趋势
    if [(x < 45 or y < 160 or z < 11.5)] != [(x < 45 or y < 160 or z * 8 < 11.5)]:
        triggered.add(119)
    if [(x < 45 or y < 160 or z < 11.5)] != [(x < 45 or y * 8 < 160 or z < 11.5)]:
        triggered.add(120)

    # 121-122：检测到参数偏高趋势
    if [(x > 55 or y > 190 or z > 13.5)] != [(x * 2 > 55 or y > 190 or z > 13.5)]:
        triggered.add(121)
    if [(x > 55 or y > 190 or z > 13.5)] != [(x > 55 or y * 8 > 190 or z > 13.5)]:
        triggered.add(122)

    # 123-124：温度电压比需要调整
    if [(abs(x / (y / 4) - 1.14) > 0.15)] != [(abs(x / (y / 4) - 1.14) > 15)]:
        triggered.add(123)
    if [(abs(x / (y / 4) - 1.14) > 0.15)] != [(abs(x / (y / 4) - 1.14) > 10)]:
        triggered.add(124)

    # 125-126：温度流量比需要调整
    if [(abs(x / z - 4.0) > 0.5)] != [(abs(x / z - 4.0) > 5)]:
        triggered.add(125)
    if [(abs(x / z - 4.0) > 0.5)] != [(abs(x / z - 4.0) > 21)]:
        triggered.add(126)

    # 127-128：电压流量比需要调整
    if [(abs(y / z - 14) > 2)] != [(abs(y / z - 14) > 99)]:
        triggered.add(127)
    if [(abs(y / z - 14) > 2)] != [(abs(y / z - 14) > 7)]:
        triggered.add(128)

    # 129-130：综合指标需要调整
    if [((x + y + z * 10) < 270 or (x + y + z * 10) > 290)] != [((x + y + z * 10) < 270 or (x + y + z * 10) > 490)]:
        triggered.add(129)
    if [((x + y + z * 10) < 270 or (x + y + z * 10) > 290)] != [((x + y + z * 10) < 270 or (x + y + z * 10) > 890)]:
        triggered.add(130)

    # 131-132：三元积需要调整
    if [((x * y * z) < 22000 or (x * y * z) > 28000)] != [((x * y * z) < 22000 or (x * y * z) > 98000)]:
        triggered.add(131)
    if [((x * y * z) < 22000 or (x * y * z) > 28000)] != [((x * y * z) < 22000 or (x * y * z) > 99000)]:
        triggered.add(132)

    # 133-134：温度电压同时偏低
    if [(x < 43 and y < 160)] != [(x < 43 and y < 1680)]:
        triggered.add(133)
    if [(x < 43 and y < 160)] != [(x < 43 and y < 1600)]:
        triggered.add(134)

    # 135-136：温度电压同时偏高
    if [(x > 57 and y > 190)] != [(x > 57 and y * 9 > 190)]:
        triggered.add(135)
    if [(x > 57 and y > 190)] != [(x > 57 and y * 8 > 190)]:
        triggered.add(136)

    # 137-138：温度流量同时偏低
    if [(x < 43 and z < 11)] != [(x < 43 and z * 2 < 11)]:
        triggered.add(137)
    if [(x < 43 and z < 11)] != [(x < 43 and 8 < 11)]:
        triggered.add(138)

    # 139-140：温度流量同时偏高
    if [(x > 57 and z > 14)] != [(x > 57 and z * 8 > 14)]:
        triggered.add(139)
    if [(x > 57 and z > 14)] != [(x > 57 and 59 > 14)]:
        triggered.add(140)

    # 141-142：电压流量同时偏低
    if [(y < 160 and z < 11)] != [(y < 160 and z * 8 < 11)]:
        triggered.add(141)
    if [(y < 160 and z < 11)] != [(y < 160 and 9 < 11)]:
        triggered.add(142)

    # 143-144：电压流量同时偏高
    if [(y > 190 and z > 14)] != [(y > 190 and z * 44 > 14)]:
        triggered.add(143)
    if [(y > 190 and z > 14)] != [(y > 190 and 23 > 14)]:
        triggered.add(144)

    # 145-146：温度在临界范围
    if [(x < 38 or x > 62)] != [(x < 38 or x * 78 > 62)]:
        triggered.add(145)
    if [(x < 38 or x > 62)] != [(x < 38 or x > 162)]:
        triggered.add(146)

    # 147-148：电压在临界范围
    if [(y < 145 or y > 205)] != [(y < 145 or y > 150)]:
        triggered.add(147)
    if [(y < 145 or y > 205)] != [(y < 115 or y > 205)]:
        triggered.add(148)

    # 149-150：流量在临界范围
    if [(z < 9.5 or z > 15.5)] != [(z < 9.5 or z > 155)]:
        triggered.add(149)
    if [(z < 9.5 or z > 15.5)] != [(z < 9.5 or z > 95.5)]:
        triggered.add(150)

    # 151-152：所有参数严重偏低
    if [(x < 35 and y < 155 and z < 10.5)] != [(x < 35 and y < 155 and z < 105)]:
        triggered.add(151)
    if [(x < 35 and y < 155 and z < 10.5)] != [(x < 35 and y < 155 and z * 8 < 10.5)]:
        triggered.add(152)

    # 153-154：三维偏离过大
    if [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 16) > 100] != [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 16) > 900]:
        triggered.add(153)
    if [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 16) > 100] != [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 89) > 100]:
        triggered.add(154)

    # 155-156：三元积偏离过大
    if [(abs((x * y * z) / 25000 - 1) > 0.2)] != [(abs((x * y * z) / 25000 - 1) > 2)]:
        triggered.add(155)
    if [(abs((x * y * z) / 25000 - 1) > 0.2)] != [(abs((x * y * z) / 25000 - 1) > 7)]:
        triggered.add(156)

    return triggered


# === 绑定规则函数（使用第10类） ===
execute_Tr = section10_comprehensive_hybrid_control

# === 状态处理辅助函数（使用新范围） ===
def clip_state(state):
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return np.clip(state, min_vals, max_vals)

def normalize_state(state):
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return 2 * (state - min_vals) / (max_vals - min_vals) - 1

def denormalize_state(normalized_state):
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return (normalized_state + 1) * (max_vals - min_vals) / 2 + min_vals

def coverage_similarity(triggered, target_path):
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0
    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)

def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']
    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']
    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']
    return reward

# === SAC Actor（动作步长针对新范围调整） ===
class GaussianPolicy(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dim=256):
        super(GaussianPolicy, self).__init__()
        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.mean = nn.Linear(hidden_dim, action_dim)
        self.log_std = nn.Linear(hidden_dim, action_dim)
        # 动作步长：温度±4，电压±6，流量±1.5（约为各维度范围的6%~7.5%）
        self.register_buffer('action_scale', torch.tensor([4.0, 6.0, 1.5]))
        self.register_buffer('action_bias', torch.tensor([0.0, 0.0, 0.0]))

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))
        mean = self.mean(x)
        log_std = self.log_std(x)
        log_std = torch.clamp(log_std, min=-20, max=2)
        return mean, log_std

    def sample(self, state):
        mean, log_std = self.forward(state)
        std = log_std.exp()
        normal = Normal(mean, std)
        x_t = normal.rsample()
        action = torch.tanh(x_t) * self.action_scale + self.action_bias
        log_prob = normal.log_prob(x_t)
        log_prob -= torch.log(self.action_scale * (1 - torch.tanh(x_t).pow(2)) + 1e-6)
        log_prob = log_prob.sum(1, keepdim=True)
        mean = torch.tanh(mean) * self.action_scale + self.action_bias
        return action, log_prob, mean

# === SAC Critic ===
class QNetwork(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dim=256):
        super(QNetwork, self).__init__()
        self.fc1 = nn.Linear(state_dim + action_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, 1)
        self.fc4 = nn.Linear(state_dim + action_dim, hidden_dim)
        self.fc5 = nn.Linear(hidden_dim, hidden_dim)
        self.fc6 = nn.Linear(hidden_dim, 1)

    def forward(self, state, action):
        xu = torch.cat([state, action], 1)
        x1 = F.relu(self.fc1(xu))
        x1 = F.relu(self.fc2(x1))
        q1 = self.fc3(x1)
        x2 = F.relu(self.fc4(xu))
        x2 = F.relu(self.fc5(x2))
        q2 = self.fc6(x2)
        return q1, q2

# === 经验回放缓冲区 ===
class EnhancedReplayBuffer:
    def __init__(self, capacity=50000):
        self.buffer = deque(maxlen=capacity)
        self.experience_info = deque(maxlen=capacity)

    def push(self, state, action, reward, next_state, done, path_idx, similarity):
        self.buffer.append((state, action, reward, next_state, done))
        self.experience_info.append({'path_idx': path_idx, 'similarity': similarity})

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return None
        batch = random.sample(self.buffer, batch_size)
        state, action, reward, next_state, done = map(np.stack, zip(*batch))
        return (
            torch.FloatTensor(state).to(device),
            torch.FloatTensor(action).to(device),
            torch.FloatTensor(reward).unsqueeze(1).to(device),
            torch.FloatTensor(next_state).to(device),
            torch.FloatTensor(done).unsqueeze(1).to(device)
        )

    def get_top_k_per_path(self, num_paths, k=20):
        path_samples = {i: [] for i in range(num_paths)}
        for idx, info in enumerate(self.experience_info):
            path_idx = info['path_idx']
            similarity = info['similarity']
            path_samples[path_idx].append((idx, similarity, self.buffer[idx]))

        top_k_results = {}
        for path_idx in range(num_paths):
            samples = path_samples[path_idx]
            if len(samples) == 0:
                top_k_results[path_idx] = []
                continue
            samples.sort(key=lambda x: x[1], reverse=True)
            top_k = samples[:k]
            top_k_results[path_idx] = []
            for sample in top_k:
                normalized_state = sample[2][0]
                original_state = denormalize_state(normalized_state)
                x, y, z = round(original_state[0]), round(original_state[1]), round(original_state[2], 1)
                triggered = execute_Tr(x, y, z)
                top_k_results[path_idx].append({
                    'state': np.array([x, y, z], dtype=np.float32),
                    'similarity': sample[1],
                    'triggered': triggered
                })
        return top_k_results

    def __len__(self):
        return len(self.buffer)

# === SAC Agent ===
class SACAgent:
    def __init__(self, state_dim=3, action_dim=3):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.policy = GaussianPolicy(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.policy_optimizer = optim.Adam(self.policy.parameters(), lr=EXPERIMENT_CONFIG['ACTOR_LR'])
        self.critic = QNetwork(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_target = QNetwork(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_target.load_state_dict(self.critic.state_dict())
        self.critic_optimizer = optim.Adam(self.critic.parameters(), lr=EXPERIMENT_CONFIG['CRITIC_LR'])
        self.target_entropy = -action_dim
        self.log_alpha = torch.zeros(1, requires_grad=True, device=device)
        self.alpha_optimizer = optim.Adam([self.log_alpha], lr=EXPERIMENT_CONFIG['ALPHA_LR'])
        self.replay_buffer = EnhancedReplayBuffer()
        self.replay_train_count = 0

    def get_action(self, state, deterministic=False):
        normalized_state = normalize_state(state)
        state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)
        with torch.no_grad():
            if deterministic:
                _, _, action = self.policy.sample(state_tensor)
            else:
                action, _, _ = self.policy.sample(state_tensor)
        action = action.cpu().numpy()[0]
        return action

    def store_experience(self, state, action, reward, next_state, done, path_idx, similarity):
        normalized_state = normalize_state(state)
        normalized_next_state = normalize_state(next_state)
        self.replay_buffer.push(
            normalized_state, action, reward,
            normalized_next_state, done, path_idx, similarity
        )

    def replay_train(self):
        batch = self.replay_buffer.sample(EXPERIMENT_CONFIG['REPLAY_BATCH_SIZE'])
        if batch is None:
            return
        state, action, reward, next_state, done = batch

        with torch.no_grad():
            next_action, next_log_prob, _ = self.policy.sample(next_state)
            q1_next, q2_next = self.critic_target(next_state, next_action)
            q_next = torch.min(q1_next, q2_next)
            target_q = reward + (1 - done) * EXPERIMENT_CONFIG['GAMMA'] * (
                    q_next - self.log_alpha.exp() * next_log_prob)

        q1, q2 = self.critic(state, action)
        critic_loss = F.mse_loss(q1, target_q) + F.mse_loss(q2, target_q)
        self.critic_optimizer.zero_grad()
        critic_loss.backward()
        self.critic_optimizer.step()

        new_action, log_prob, _ = self.policy.sample(state)
        q1_new, q2_new = self.critic(state, new_action)
        q_new = torch.min(q1_new, q2_new)
        policy_loss = (self.log_alpha.exp() * log_prob - q_new).mean()
        self.policy_optimizer.zero_grad()
        policy_loss.backward()
        self.policy_optimizer.step()

        alpha_loss = -(self.log_alpha * (log_prob + self.target_entropy).detach()).mean()
        self.alpha_optimizer.zero_grad()
        alpha_loss.backward()
        self.alpha_optimizer.step()

        for param, target_param in zip(self.critic.parameters(), self.critic_target.parameters()):
            target_param.data.copy_(
                EXPERIMENT_CONFIG['TAU'] * param.data + (1 - EXPERIMENT_CONFIG['TAU']) * target_param.data)

        self.replay_train_count += 1
        if self.replay_train_count % 2 == 0:
            alpha_value = self.log_alpha.exp().item()
            print(f"  -> SAC 更新 (第 {self.replay_train_count} 次), Alpha={alpha_value:.4f}")

# === 性能计算函数 ===
def calculate_run_performance(run_idx, sac_results, training_time, total_steps, update_count, agent):
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)
    all_similarities = []
    total_samples = 0
    total_reward = 0

    for path_idx in range(num_paths):
        samples = sac_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']
            total_reward += reward
            all_similarities.append(similarity)
            total_samples += 1

    average_reward = total_reward / total_samples if total_samples > 0 else 0
    convergence = np.mean(all_similarities) if all_similarities else 0
    environment_adaptability = 1 / (np.std(all_similarities) + 1e-8) if len(all_similarities) > 1 else 0
    generalization_ability = convergence
    computational_efficiency = total_steps / training_time if training_time > 0 else 0
    policy_update_frequency = update_count / training_time if training_time > 0 else 0
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0

    return {
        'Run': run_idx + 1,
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }

# === Excel 导出函数（与之前相同） ===
def export_to_excel(all_sac_results, all_performance_data, target_paths, output_path="SAC_20_run.xlsx"):
    print("\n正在导出数据到 Excel...")
    all_sac_summary_data = []
    all_sac_detailed_data = []

    for run_idx, (sac_results, performance_data) in enumerate(zip(all_sac_results, all_performance_data)):
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = sac_results[path_idx]
            if len(samples) == 0:
                all_sac_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue
            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = 'Yes' if perfect_count > 0 else 'No'
            all_sac_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })

            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']
                all_sac_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Sample ID': sample_idx + 1,
                    'X(温度)': int(state[0]),
                    'Y(电压)': int(state[1]),
                    'Z(流量)': float(state[2]),
                    'Similarity': round(similarity, 4),
                    'Perfect Coverage': 'Yes' if abs(similarity - 1.0) < 0.001 else 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path))),
                    'Triggered Rules': ', '.join(map(str, sorted(triggered))),
                    'Hit Rule Count': len(target_path.intersection(triggered)),
                    'Target Rule Count': len(target_path)
                })

    summary_df = pd.DataFrame(all_sac_summary_data)
    detailed_df = pd.DataFrame(all_sac_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='SACPath', index=False)
        detailed_df.to_excel(writer, sheet_name='SACDetailed Sample Data', index=False)
        selected_columns = [
            'Run', 'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df[selected_columns].to_excel(writer, sheet_name='Metric', index=False)

        workbook = writer.book
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        perfect_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        for sheet_name in ['SACPath', 'SACDetailed Sample Data', 'Metric']:
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

        ws1 = writer.sheets['SACPath']
        for row_idx in range(2, ws1.max_row + 1):
            if ws1.cell(row_idx, 9).value == 'Yes':
                for col_idx in range(1, ws1.max_column + 1):
                    ws1.cell(row_idx, col_idx).fill = perfect_fill

    print(f"文件已保存到: {output_path}")

# === 单次训练流程 ===
def train_sac_workflow():
    print("=" * 80)
    print("SAC 训练 - 温度/电压/流量混合控制 (第10类规则)")
    print("状态范围: 温度 1~65, 电压 100~205, 流量 1~20")
    print("相似度 = 命中规则数 / 目标路径规则数")
    print("=" * 80)

    agent = SACAgent()
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n生成初始样本: {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']} / 路径")
    path_samples = {}
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']

    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            state = np.random.uniform(min_vals, max_vals).astype(np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  路径 {path_idx + 1}/{num_paths}: {len(samples)} 个初始样本")

    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size

    print(f"\n批次大小: {batch_size}, 单样本迭代步数: {EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(f"总批次数: {num_batches} /路径 × {num_paths} 路径 = {num_batches * num_paths}")
    print("-" * 80)

    for batch_idx in range(num_batches):
        print(f"\n批次 {batch_idx + 1}/{num_batches}")

        for path_idx in range(num_paths):
            target_path = target_paths[path_idx]
            batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]

            batch_rewards = []
            batch_similarities = []

            for initial_state in batch_samples:
                state = initial_state.copy()
                episode_reward = 0
                final_similarity = 0

                for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                    action = agent.get_action(state, deterministic=False)
                    next_state = state + action
                    next_state = clip_state(next_state)

                    x, y, z = next_state
                    triggered = execute_Tr(x, y, z)
                    reward = unified_reward_function(triggered, target_path)
                    similarity = coverage_similarity(triggered, target_path)

                    done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)

                    agent.store_experience(
                        state, action, reward, next_state, done,
                        path_idx, similarity
                    )

                    state = next_state
                    episode_reward += reward
                    final_similarity = similarity
                    total_steps += 1

                batch_rewards.append(episode_reward)
                batch_similarities.append(final_similarity)

            avg_reward = np.mean(batch_rewards)
            avg_sim = np.mean(batch_similarities)
            print(f"  路径 {path_idx+1}: 平均奖励={avg_reward:.2f}, 平均相似度={avg_sim:.4f}")

        print("\n  执行SAC网络参数更新...")
        agent.replay_train()

    training_time = time.time() - start_time

    print("\n" + "=" * 80)
    print(f"本轮训练完成 | 总耗时: {training_time:.2f}s | 总交互步数: {total_steps}")
    print(f"网络更新次数: {agent.replay_train_count}")
    print("=" * 80)

    print(f"\n提取每条路径 Top-{EXPERIMENT_CONFIG['TOP_K_SAMPLES']} 高相似度样本...")
    top_k_results = agent.replay_buffer.get_top_k_per_path(num_paths, EXPERIMENT_CONFIG['TOP_K_SAMPLES'])

    return agent, top_k_results, training_time, total_steps, agent.replay_train_count

# === 主程序入口 ===
def main():
    print("=" * 80)
    print("SAC 温度/电压/流量混合控制 (第10类规则) - 20轮完整实验")
    print("=" * 80)

    all_sac_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'='*80}")
        print(f"第 {run_idx+1}/{EXPERIMENT_CONFIG['NUM_RUNS']} 轮实验启动")
        print(f"{'='*80}")

        agent, sac_results, training_time, total_steps, update_count = train_sac_workflow()
        performance = calculate_run_performance(
            run_idx, sac_results, training_time, total_steps, update_count, agent
        )

        all_sac_results.append(sac_results)
        all_performance_data.append(performance)

        print(f"\n第 {run_idx+1} 轮核心指标:")
        print(f"  平均相似度: {performance['Average Similarity']:.4f}")
        print(f"  最高相似度: {performance['Max Similarity']:.4f}")
        print(f"  平均奖励: {performance['Average Reward']:.2f}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"SAC_section10_20run_{timestamp}.xlsx"
    export_to_excel(all_sac_results, all_performance_data, target_paths, output_file)

    avg_sim_list = [p['Average Similarity'] for p in all_performance_data]
    max_sim_list = [p['Max Similarity'] for p in all_performance_data]
    print("\n" + "=" * 80)
    print("20轮实验整体汇总:")
    print(f"  平均相似度 均值: {np.mean(avg_sim_list):.4f}")
    print(f"  平均相似度 标准差: {np.std(avg_sim_list):.4f}")
    print(f"  单轮最高平均相似度: {np.max(avg_sim_list):.4f}")
    print(f"  全局最高相似度: {np.max(max_sim_list):.4f}")
    print("=" * 80)


if __name__ == "__main__":
    main()