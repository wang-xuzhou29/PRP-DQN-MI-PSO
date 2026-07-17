import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
import numpy as np
import random
import time
import math
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
from datetime import datetime

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# === 目标路径组（与 section10 规则编号 1~156 匹配） ===
targetPaths = [
    {1, 2, 3, 4, 5, 6, 11, 12, 17, 21, 22, 23, 24, 29, 31, 32, 33, 34, 36, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48, 53,
     54, 55, 56, 67, 68, 69, 70, 71, 72, 77, 78, 79, 80, 95, 96, 97, 99, 100, 101, 102, 103, 105, 106, 113, 114, 117,
     118, 123, 124, 125, 126, 129, 130, 135, 136, 139, 140, 146, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 17, 22, 23, 29, 31, 32, 33, 34, 36, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48, 53, 54, 55, 56, 57,
     58, 59, 60, 67, 68, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 101, 102, 105, 106, 113, 114, 117, 118, 123, 124,
     125, 126, 127, 128, 129, 130, 135, 136, 139, 140, 146, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 11, 12, 17, 19, 21, 22, 23, 24, 29, 31, 32, 33, 34, 36, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48,
     53, 54, 55, 56, 67, 68, 69, 70, 77, 78, 79, 80, 95, 96, 97, 99, 100, 101, 102, 103, 105, 106, 113, 114, 117, 118,
     123, 124, 125, 126, 129, 130, 135, 136, 139, 140, 145, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 11, 12, 15, 16, 17, 23, 24, 29, 31, 32, 33, 34, 36, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48, 53,
     54, 55, 56, 59, 60, 67, 68, 69, 70, 77, 78, 79, 80, 95, 96, 97, 99, 100, 101, 102, 105, 106, 113, 114, 117, 118,
     123, 124, 125, 126, 129, 130, 135, 136, 139, 140, 146, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 11, 12, 17, 22, 23, 24, 29, 31, 33, 34, 36, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48, 53, 54, 55,
     56, 67, 68, 69, 70, 71, 72, 77, 78, 79, 80, 95, 96, 97, 99, 100, 101, 102, 103, 105, 106, 113, 114, 117, 118, 123,
     124, 125, 126, 129, 130, 139, 140, 143, 144, 146, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 11, 12, 17, 22, 23, 24, 29, 31, 32, 33, 34, 35, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48, 53, 54,
     55, 56, 59, 60, 67, 68, 69, 70, 77, 78, 79, 80, 95, 96, 97, 101, 102, 105, 106, 113, 114, 117, 118, 120, 123, 124,
     125, 126, 129, 130, 135, 136, 139, 140, 142, 146, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 17, 22, 23, 29, 31, 33, 34, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48, 51, 52, 53, 54, 55, 56, 69,
     70, 71, 72, 73, 74, 77, 78, 79, 80, 95, 96, 97, 99, 100, 101, 102, 103, 105, 106, 108, 113, 114, 117, 118, 123,
     124, 125, 126, 129, 130, 139, 140, 143, 144, 146, 147, 153, 156},

    {1, 2, 5, 6, 9, 10, 17, 19, 20, 22, 23, 24, 28, 29, 35, 37, 38, 41, 42, 49, 53, 54, 55, 56, 57, 58, 63, 64, 66, 68,
     71, 72, 77, 78, 79, 80, 95, 96, 101, 102, 105, 106, 109, 113, 114, 117, 118, 120, 121, 122, 123, 124, 125, 126,
     127, 128, 129, 130, 131, 132, 142, 145, 148, 153, 155, 156},

    {1, 2, 3, 4, 5, 6, 17, 19, 22, 23, 29, 33, 34, 35, 37, 38, 39, 40, 41, 42, 53, 54, 55, 56, 57, 58, 59, 60, 67, 68,
     69, 70, 73, 74, 77, 78, 79, 80, 85, 86, 95, 96, 97, 101, 102, 105, 106, 113, 114, 117, 118, 120, 123, 124, 127,
     128, 129, 130, 135, 136, 139, 140, 142, 145, 148, 153, 156},

    {1, 2, 3, 4, 17, 19, 21, 22, 29, 31, 32, 33, 34, 37, 38, 39, 40, 47, 48, 51, 52, 53, 54, 55, 56, 57, 58, 63, 64, 69,
     70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 99, 100, 101, 102, 105, 106, 113, 114, 117, 118, 123, 124, 125, 126, 127,
     128, 129, 130, 135, 136, 145, 147, 149, 150, 153, 156},

    {1, 2, 3, 4, 5, 6, 18, 19, 22, 23, 25, 26, 29, 31, 36, 37, 38, 39, 40, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 55,
     56, 57, 58, 68, 69, 70, 71, 72, 77, 78, 79, 80, 84, 97, 101, 102, 103, 105, 106, 108, 117, 118, 119, 125, 126, 127,
     128, 129, 130, 139, 140, 143, 144, 145, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 22, 23, 25, 26, 29, 31, 35, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 53, 54, 56, 57, 58,
     61, 62, 68, 69, 70, 71, 72, 77, 78, 79, 80, 97, 101, 102, 105, 106, 108, 117, 118, 119, 123, 124, 125, 126, 127,
     128, 129, 130, 139, 140, 143, 144, 146, 147, 153, 156},

    {1, 2, 3, 4, 17, 19, 22, 29, 33, 34, 37, 38, 39, 40, 41, 42, 53, 54, 55, 56, 57, 58, 59, 60, 63, 64, 69, 70, 73, 74,
     77, 78, 79, 80, 87, 88, 95, 96, 97, 101, 102, 105, 106, 113, 114, 117, 118, 120, 123, 124, 125, 126, 127, 128, 129,
     130, 135, 136, 142, 145, 148, 149, 150, 153, 156},

    {1, 2, 3, 4, 5, 6, 18, 22, 23, 24, 29, 31, 36, 37, 38, 39, 40, 43, 44, 45, 46, 47, 48, 51, 52, 53, 54, 55, 56, 57,
     58, 67, 68, 69, 70, 71, 72, 77, 78, 79, 80, 83, 84, 97, 101, 102, 103, 105, 106, 108, 117, 118, 125, 126, 127, 128,
     129, 130, 139, 140, 143, 144, 145, 147, 153, 156},

    {1, 2, 5, 6, 17, 19, 20, 22, 23, 27, 29, 35, 37, 38, 39, 40, 41, 42, 49, 50, 53, 54, 55, 56, 57, 58, 63, 64, 66, 67,
     68, 69, 70, 71, 72, 77, 78, 79, 80, 95, 96, 101, 102, 105, 106, 113, 114, 117, 118, 120, 123, 124, 125, 126, 127,
     128, 129, 130, 131, 132, 142, 145, 148, 153, 156},

    {1, 2, 3, 4, 13, 14, 17, 22, 29, 31, 32, 33, 34, 37, 38, 39, 40, 51, 52, 53, 54, 55, 56, 57, 58, 63, 64, 69, 70, 75,
     76, 77, 78, 79, 80, 89, 90, 93, 94, 95, 96, 97, 99, 100, 101, 102, 105, 106, 112, 113, 114, 117, 118, 125, 126,
     127, 128, 129, 130, 145, 147, 149, 150, 153, 156},

    {1, 2, 3, 4, 7, 8, 13, 14, 17, 21, 22, 27, 29, 33, 34, 37, 38, 39, 40, 51, 52, 56, 57, 58, 63, 64, 69, 70, 77, 78,
     79, 80, 87, 88, 89, 90, 92, 93, 94, 95, 96, 97, 99, 100, 101, 102, 105, 106, 111, 112, 113, 114, 117, 118, 125,
     126, 127, 128, 129, 130, 145, 147, 149, 150, 156},

    {1, 2, 5, 6, 9, 10, 17, 19, 20, 23, 24, 28, 30, 35, 41, 42, 49, 53, 54, 55, 56, 57, 58, 63, 64, 66, 68, 71, 72, 77,
     78, 79, 80, 95, 96, 101, 102, 105, 106, 109, 110, 113, 114, 117, 118, 121, 122, 123, 124, 125, 126, 127, 128, 129,
     130, 131, 132, 142, 145, 148, 153, 155, 156},

    {1, 2, 3, 4, 5, 6, 17, 22, 23, 24, 28, 29, 35, 37, 38, 50, 53, 54, 56, 57, 58, 63, 64, 66, 68, 71, 72, 77, 78, 79,
     80, 89, 90, 95, 96, 101, 102, 105, 106, 109, 113, 114, 117, 118, 121, 122, 123, 124, 125, 126, 127, 128, 129, 130,
     131, 132, 138, 142, 148, 151, 153, 155, 156},

    {1, 2, 3, 4, 5, 6, 11, 12, 15, 16, 17, 19, 23, 24, 28, 29, 36, 37, 38, 39, 40, 49, 53, 54, 55, 56, 63, 64, 69, 70,
     79, 80, 82, 89, 90, 95, 96, 101, 102, 103, 105, 106, 113, 114, 116, 117, 118, 121, 122, 123, 124, 125, 126, 129,
     130, 131, 132, 133, 134, 138, 145, 147, 156},

    {1, 2, 3, 4, 5, 6, 17, 19, 20, 22, 23, 27, 29, 37, 38, 39, 40, 49, 50, 51, 52, 53, 54, 55, 56, 63, 64, 67, 68, 69,
     70, 77, 78, 79, 80, 87, 88, 89, 90, 95, 96, 97, 101, 102, 103, 105, 106, 113, 114, 116, 117, 118, 123, 124, 125,
     126, 129, 130, 143, 144, 145, 147, 154, 156},

    {1, 2, 3, 4, 7, 8, 19, 20, 21, 22, 25, 26, 28, 30, 35, 39, 40, 43, 44, 46, 53, 54, 55, 56, 57, 58, 61, 62, 65, 66,
     73, 74, 77, 78, 79, 80, 101, 102, 104, 105, 106, 109, 110, 117, 118, 119, 121, 122, 123, 124, 125, 126, 127, 129,
     130, 131, 132, 145, 147, 153, 155, 156},

    {1, 2, 3, 4, 5, 6, 19, 21, 22, 28, 29, 35, 37, 38, 39, 40, 43, 44, 45, 46, 53, 54, 57, 58, 61, 62, 66, 73, 74, 77,
     78, 79, 80, 101, 102, 103, 104, 105, 106, 117, 118, 121, 122, 123, 124, 125, 126, 127, 128, 129, 130, 131, 132,
     133, 134, 137, 145, 147, 153, 155, 156},

    {1, 2, 5, 6, 17, 23, 29, 33, 34, 35, 37, 38, 41, 42, 43, 44, 53, 54, 55, 56, 57, 58, 59, 60, 67, 68, 69, 70, 73, 74,
     79, 80, 95, 96, 98, 101, 102, 105, 106, 109, 113, 114, 117, 118, 120, 123, 124, 127, 128, 129, 130, 131, 132, 135,
     136, 139, 140, 142, 146, 153, 156},

    {1, 2, 3, 4, 5, 6, 22, 25, 26, 29, 31, 32, 35, 37, 38, 41, 42, 43, 44, 45, 46, 47, 48, 50, 53, 54, 56, 57, 58, 59,
     60, 69, 70, 79, 80, 101, 102, 105, 106, 109, 117, 118, 123, 124, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136,
     139, 140, 141, 146, 147, 153, 156},

    {1, 2, 17, 19, 27, 29, 35, 37, 38, 39, 40, 41, 42, 49, 50, 53, 54, 56, 57, 58, 63, 64, 66, 67, 68, 72, 77, 78, 79,
     80, 95, 96, 101, 102, 105, 106, 107, 109, 113, 114, 117, 118, 123, 124, 125, 126, 127, 129, 130, 131, 132, 138,
     142, 145, 149, 150, 153, 155, 156},

    {1, 2, 3, 4, 5, 6, 7, 8, 13, 14, 15, 16, 23, 25, 26, 28, 29, 35, 37, 38, 39, 40, 43, 44, 45, 46, 55, 56, 57, 58, 66,
     69, 70, 77, 78, 79, 80, 81, 92, 103, 104, 105, 106, 117, 118, 119, 121, 122, 125, 126, 127, 128, 129, 130, 131,
     132, 145, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 17, 19, 20, 21, 22, 23, 24, 28, 29, 36, 37, 38, 39, 40, 49, 50, 53, 54, 67,
     68, 69, 70, 87, 88, 95, 96, 98, 101, 102, 103, 105, 106, 113, 114, 115, 116, 117, 118, 121, 122, 129, 130, 145,
     147, 156},

    {1, 2, 3, 4, 5, 6, 7, 8, 13, 14, 17, 23, 27, 29, 36, 37, 38, 39, 40, 49, 50, 63, 64, 67, 68, 69, 70, 79, 80, 84, 85,
     86, 87, 88, 89, 90, 91, 92, 95, 96, 97, 101, 102, 103, 105, 106, 113, 114, 116, 117, 118, 129, 130, 145, 147, 154,
     156},

    {3, 4, 5, 6, 22, 28, 30, 35, 45, 46, 53, 54, 55, 56, 63, 64, 65, 66, 71, 72, 73, 74, 77, 78, 79, 80, 103, 109, 110,
     117, 118, 121, 122, 123, 124, 125, 126, 127, 128, 131, 132, 137, 141, 147, 152, 153, 155, 156},

]

# === 配置（新范围：X: 1~65, Y: 100~205, Z: 1~20） ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'ACTION_DIM': 3,
    'MIN_VALUES': np.array([1.0, 100.0, 1.0]),
    'MAX_VALUES': np.array([65.0, 205.0, 20.0]),
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 3,
    'PPO_BATCH_SIZE': 64,
    'PPO_EPOCHS': 4,
    'CLIP_EPSILON': 0.2,
    'GAMMA': 0.99,
    'GAE_LAMBDA': 0.95,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'ACTOR_LR': 3e-4,
    'CRITIC_LR': 3e-4,
    'NUM_RUNS': 20,
    'TOP_K_SAMPLES': 20,
    'TARGET_PATHS': targetPaths
}

# ==================== section10 触发函数（规则编号 1~156） ====================
def section10_comprehensive_hybrid_control(x, y, z):
    """第10类: 综合混合控制 (湿度x, 扭矩y, 电流z) - 156个有效分支（删除99、100后顺延编号）"""
    triggered = set()

    # -------------------------- 1-40：综合系统基础检查 --------------------------
    if [(30 < x < 70)] != [(30 < x * 78 < 70)]:
        triggered.add(1)
    if [(30 < x < 70)] != [(30 < 122 < 70)]:
        triggered.add(2)
    if [(140 < y < 210)] != [(140 < 678 < 210)]:
        triggered.add(3)
    if [(140 < y < 210)] != [(140 < y * 88 < 210)]:
        triggered.add(4)
    if [(8 < z < 16)] != [(8 < z * 67 < 16)]:
        triggered.add(5)
    if [(8 < z < 16)] != [(8 < 566 < 16)]:
        triggered.add(6)
    if [(45 < x < 55 and 160 < y < 190)] != [(45 < x < 55 and 160 < y * 8 < 190)]:
        triggered.add(7)
    if [(45 < x < 55 and 160 < y < 190)] != [(45 < x * 6 < 55 and 160 < y < 190)]:
        triggered.add(8)
    if [(42 < x < 58 and 11 < z < 14)] != [(42 < x < 58 and 11 < z * 78 < 14)]:
        triggered.add(9)
    if [(42 < x < 58 and 11 < z < 14)] != [(42 < x < 58 and 11 < z * 8 < 14)]:
        triggered.add(10)
    if [(155 < y < 195 and 11 < z < 14)] != [(155 < y < 195 and 11 < z * 78 < 14)]:
        triggered.add(11)
    if [(155 < y < 195 and 11 < z < 14)] != [(155 < y * 66 < 195 and 11 < z < 14)]:
        triggered.add(12)
    if [(48 < x < 52)] != [(48 < x * 78 < 52)]:
        triggered.add(13)
    if [(48 < x < 52)] != [(48 < 88 < 52)]:
        triggered.add(14)
    if [(173 < y < 177)] != [(173 < y * 7 < 177)]:
        triggered.add(15)
    if [(173 < y < 177)] != [(173 < 88 < 177)]:
        triggered.add(16)
    if [(12.2 < z < 12.8)] != [(12.2 < z < 128)]:
        triggered.add(17)
    if [(12.2 < z < 12.8)] != [(10 < z < 12.8)]:
        triggered.add(18)
    if [(abs(x - 50) < 2)] != [(abs(x - 50) < 12)]:
        triggered.add(19)
    if [(abs(x - 50) < 2)] != [(abs(x - 50) < 8)]:
        triggered.add(20)
    if [(abs(y - 175) < 5)] != [(abs(y - 175) < 15)]:
        triggered.add(21)
    if [(abs(y - 175) < 5)] != [(abs(y - 175) < 51)]:
        triggered.add(22)
    if [(abs(z - 12.5) < 0.3)] != [(abs(z - 12.5) < 3)]:
        triggered.add(23)
    if [(abs(z - 12.5) < 0.3)] != [(abs(z - 12.5) < 1)]:
        triggered.add(24)
    if [(x > 42 and y > 155 and z > 11)] != [(x > 42 and y > 155 and 22 > 11)]:
        triggered.add(25)
    if [(x > 42 and y > 155 and z > 11)] != [(x > 42 and y > 155 and z * 3 > 11)]:
        triggered.add(26)
    if [(x < 58 and y < 195 and z < 14)] != [(x < 58 and y < 195 and 10 < 14)]:
        triggered.add(27)
    if [(x < 58 and y < 195 and z < 14)] != [(x < 58 and y < 195 and z * 78 < 14)]:
        triggered.add(28)
    if [(260 < x + y + z * 10 < 300)] != [(260 < x + y + z * 10 < 3100)]:
        triggered.add(29)
    if [(260 < x + y + z * 10 < 300)] != [(300 < x + y + z * 10 < 300)]:
        triggered.add(30)
    if [(8000 < x * y < 10000)] != [(8000 < x * y < 70000)]:
        triggered.add(31)
    if [(8000 < x * y < 10000)] != [(8000 < x * y < 12000)]:
        triggered.add(32)
    if [(600 < x * z < 750)] != [(600 < x * z < 7510)]:
        triggered.add(33)
    if [(600 < x * z < 750)] != [(600 < x * z < 1750)]:
        triggered.add(34)
    if [(2100 < y * z < 2500)] != [(210 < y * z < 2500)]:
        triggered.add(35)
    if [(2100 < y * z < 2500)] != [(2100 < 23 * z < 2500)]:
        triggered.add(36)
    if [(86 < (x + y + z * 10) / 3 < 100)] != [(86 < (x + y + z * 10) / 3 < 1020)]:
        triggered.add(37)
    if [(86 < (x + y + z * 10) / 3 < 100)] != [(86 < (x + y + z * 10) / 3 < 200)]:
        triggered.add(38)
    if [(math.sqrt(x ** 2 + y ** 2 + (z * 10) ** 2) > 200)] != [(math.sqrt(x ** 2 + y ** 2 + (z * 10) ** 2) > 2100)]:
        triggered.add(39)
    if [(math.sqrt(x ** 2 + y ** 2 + (z * 10) ** 2) > 200)] != [(math.sqrt(x ** 2 + y ** 2 + (z * 10) ** 2) > 500)]:
        triggered.add(40)

    # -------------------------- 41-80：跨域参数分析 --------------------------
    if [(1.0 < x / (y / 4) < 1.3)] != [(1.0 < x / (y / 4) < 13)]:
        triggered.add(41)
    if [(1.0 < x / (y / 4) < 1.3)] != [(1.0 < x / (y / 4) < 333)]:
        triggered.add(42)
    if [(3.5 < x / z < 4.5)] != [(3.5 < x / z < 45)]:
        triggered.add(43)
    if [(3.5 < x / z < 4.5)] != [(3.5 < x / z < 24.5)]:
        triggered.add(44)
    if [(12 < y / z < 16)] != [(12 < y / z < 30)]:
        triggered.add(45)
    if [(12 < y / z < 16)] != [(12 < y / z < 76)]:
        triggered.add(46)
    if [(15 < (x - 40) + (y - 150) / 5 < 25)] != [(15 < (x - 40) + (y - 150) / 5 < 215)]:
        triggered.add(47)
    if [(15 < (x - 40) + (y - 150) / 5 < 25)] != [(15 < (x - 40) + (y - 150) / 5 < 125)]:
        triggered.add(48)
    if [(15 < (x - 40) + (z - 10) * 3 < 25)] != [(15 < (x - 40) + (z - 10) * 5 < 25)]:
        triggered.add(49)
    if [(15 < (x - 40) + (z - 10) * 3 < 25)] != [(15 < (x - 40) + (z - 10) * 10 < 25)]:
        triggered.add(50)
    if [(8 < (y - 150) / 5 + (z - 10) < 12)] != [(8 < (y - 150) / 5 + (z - 10) < 121)]:
        triggered.add(51)
    if [(8 < (y - 150) / 5 + (z - 10) < 12)] != [(8 < (y - 150) / 5 + (z - 10) < 112)]:
        triggered.add(52)
    if [(abs((x - 50) - (y - 175) / 4) < 5)] != [(abs((x - 50) - (y - 175) / 4) < 115)]:
        triggered.add(53)
    if [(abs((x - 50) - (y - 175) / 4) < 5)] != [(abs((x - 50) - (y - 175) / 4) < 51)]:
        triggered.add(54)
    if [(abs((x - 50) - (z - 12.5) * 3) < 6)] != [(abs((x - 50) - (z - 12.5) * 3) < 16)]:
        triggered.add(55)
    if [(abs((x - 50) - (z - 12.5) * 3) < 6)] != [(abs((x - 50) - (z - 12.5) * 3) < 26)]:
        triggered.add(56)
    if [(abs((y - 175) / 14 - (z - 12.5)) < 2)] != [(abs((y - 175) / 14 - (z - 12.5)) < 21)]:
        triggered.add(57)
    if [(abs((y - 175) / 14 - (z - 12.5)) < 2)] != [(abs((y - 175) / 14 - (z - 12.5)) < 112)]:
        triggered.add(58)
    if [(0.9 < x / (y / 4 + 10) < 1.2)] != [(0.9 < x / (y / 4 + 10) < 12)]:
        triggered.add(59)
    if [(0.9 < x / (y / 4 + 10) < 1.2)] != [(0.9 < x / (y / 4 + 10) < 112)]:
        triggered.add(60)
    if [(9 < y / (z + 5) < 13)] != [(9 < y / (z + 5) < 53)]:
        triggered.add(61)
    if [(9 < y / (z + 5) < 13)] != [(9 < y / (z + 5) < 131)]:
        triggered.add(62)
    if [(2.2 < z / (x / 10) < 2.8)] != [(2.2 < z / (x / 10) < 28)]:
        triggered.add(63)
    if [(2.2 < z / (x / 10) < 2.8)] != [(2.2 < z / (x / 10) < 128)]:
        triggered.add(64)
    if [(20000 < (x * y * z) < 30000)] != [(20000 < (x * y * z) < 56000)]:
        triggered.add(65)
    if [(20000 < (x * y * z) < 30000)] != [(20000 < (x * y * z) < 88000)]:
        triggered.add(66)
    if [(160 < x * 0.8 + y * 0.1 + z * 8 < 180)] != [(160 < x * 0.8 + y * 1 + z * 8 < 180)]:
        triggered.add(67)
    if [(160 < x * 0.8 + y * 0.1 + z * 8 < 180)] != [(160 < x * 0.8 + y * 0.1 + z * 10 < 180)]:
        triggered.add(68)
    if [((x / 50) ** 0.3 * (y / 175) ** 0.4 * (z / 12.5) ** 0.3 > 0.9)] != [((x / 50) ** 0.3 * (y / 175) ** 0.4 * (z / 12.5) ** 0.3 > 9)]:
        triggered.add(69)
    if [((x / 50) ** 0.3 * (y / 175) ** 0.4 * (z / 12.5) ** 0.3 > 0.9)] != [((x / 50) ** 0.3 * (y / 175) ** 0.4 * (z / 12.5) ** 0.3 > 19)]:
        triggered.add(70)
    if [((x - 50) * (y - 175) / 4 > -40 and (x - 50) * (y - 175) / 4 < 40)] != [((x - 50) * (y - 175) / 4 > -40 and (x - 50) * (y - 175) / 4 < 140)]:
        triggered.add(71)
    if [((x - 50) * (y - 175) / 4 > -40 and (x - 50) * (y - 175) / 4 < 40)] != [((x - 50) * (y - 175) / 4 > -40 and (x - 50) * (y - 175) / 4 < 640)]:
        triggered.add(72)
    if [((x - 50) * (z - 12.5) > -15 and (x - 50) * (z - 12.5) < 15)] != [((x - 50) * (z - 12.5) > -15 and (x - 50) * (z - 12.5) < 115)]:
        triggered.add(73)
    if [((x - 50) * (z - 12.5) > -15 and (x - 50) * (z - 12.5) < 15)] != [((x - 50) * (z - 12.5) > -15 and (x - 50) * (z - 12.5) < 151)]:
        triggered.add(74)
    if [((y - 175) / 14 * (z - 12.5) > -8 and (y - 175) / 14 * (z - 12.5) < 8)] != [((y - 175) / 14 * (z - 12.5) > -8 and (y - 175) / 14 * (z - 12.5) < 118)]:
        triggered.add(75)
    if [((y - 175) / 14 * (z - 12.5) > -8 and (y - 175) / 14 * (z - 12.5) < 8)] != [((y - 175) / 14 * (z - 12.5) > -8 and (y - 175) / 14 * (z - 12.5) < 138)]:
        triggered.add(76)
    if [(abs(x / 50 + y / 175 + z / 12.5 - 3) < 0.15)] != [(abs(x / 50 + y / 175 + z / 12.5 - 3) < 15)]:
        triggered.add(77)
    if [(abs(x / 50 + y / 175 + z / 12.5 - 3) < 0.15)] != [(abs(x / 50 + y / 175 + z / 12.5 - 3) < 325)]:
        triggered.add(78)
    if [(abs((x / 50) * (y / 175) * (z / 12.5) - 1) < 0.1)] != [(abs((x / 50) * (y / 175) * (z / 12.5) - 1) < 1)]:
        triggered.add(79)
    if [(abs((x / 50) * (y / 175) * (z / 12.5) - 1) < 0.1)] != [(abs((x / 50) * (y / 175) * (z / 12.5) - 1) < 12)]:
        triggered.add(80)

    # -------------------------- 81-156：全系统优化协调（原99、100被删除，后续顺延） --------------------------
    if [(49 < x < 51 and 173 < y < 177 and 12 < z < 17)] != [(49 < x < 51 and 173 < y < 177 and 3 < z < 17)]:
        triggered.add(81)
    if [(49 < x < 51 and 173 < y < 177 and 12 < z < 17)] != [(25 < x < 51 and 173 < y < 177 and 12 < z < 17)]:
        triggered.add(82)
    if [(abs(x / (y / 4) - 1.14) < 0.1 and abs(z - 12.5) < 0.2)] != [(abs(x / (y / 4) - 1.14) < 0.1 and abs(z - 12.5) < 1)]:
        triggered.add(83)
    if [(abs(x / (y / 4) - 1.14) < 0.1 and abs(z - 12.5) < 0.2)] != [(abs(x / (y / 4) - 1.14) < 0.1 and abs(z - 12.5) < 2)]:
        triggered.add(84)
    if [(8700 < x * y < 8800 and 12.4 < z < 12.6)] != [(8700 < x * y < 8800 and 12.4 < z < 126)]:
        triggered.add(85)
    if [(8700 < x * y < 8800 and 12.4 < z < 12.6)] != [(8700 < x * y < 8800 and 12.4 < z < 32.6)]:
        triggered.add(86)
    if [(46 < (x + y / 4) / 2 < 48 and 12.3 < z < 12.7)] != [(46 < (x + y / 4) / 2 < 48 and 12.3 < z < 127)]:
        triggered.add(87)
    if [(46 < (x + y / 4) / 2 < 48 and 12.3 < z < 12.7)] != [(46 < (x + y / 4) / 2 < 48 and 12.3 < z < 72.7)]:
        triggered.add(88)
    if [(abs(x - y / 4) < 6 and 12.3 < z < 12.7)] != [(abs(x - y / 4) < 6 and 12.3 < z < 127)]:
        triggered.add(89)
    if [(abs(x - y / 4) < 6 and 12.3 < z < 12.7)] != [(abs(x - y / 4) < 6 and 12.3 < z < 327)]:
        triggered.add(90)
    if [(math.sqrt((x - 50) ** 2 + (y / 4 - 43.75) ** 2) < 2 and abs(z - 12.5) < 0.2)] != [(math.sqrt((x - 50) ** 2 + (y / 4 - 43.75) ** 2) < 2 and abs(z - 12.5) < 2)]:
        triggered.add(91)
    if [(math.sqrt((x - 50) ** 2 + (y / 4 - 43.75) ** 2) < 2 and abs(z - 12.5) < 0.2)] != [(math.sqrt((x - 50) ** 2 + (y / 4 - 43.75) ** 2) < 2 and abs(z - 12.5) < 22)]:
        triggered.add(92)
    if [(0.98 < x / 50 < 1.02 and 0.985 < y / 175 < 1.015)] != [(0.98 < x / 50 < 1.02 and 0.985 < y / 175 < 12)]:
        triggered.add(93)
    if [(0.98 < x / 50 < 1.02 and 0.985 < y / 175 < 1.015)] != [(0.98 < x / 50 < 1.02 and 0.985 < y / 175 < 10)]:
        triggered.add(94)
    if [(0.984 < z / 12.5 < 1.016)] != [(0.984 < z / 12.5 < 10.16)]:
        triggered.add(95)
    if [(0.984 < z / 12.5 < 1.016)] != [(0.984 < z / 12.5 < 4.016)]:
        triggered.add(96)
    if [(0.985 < (x / 50 + y / 175 + z / 12.5) / 3 < 1.015)] != [(0.985 < (x / 50 + y / 175 + z / 12.5) / 3 < 10.15)]:
        triggered.add(97)
    if [(0.985 < (x / 50 + y / 175 + z / 12.5) / 3 < 1.015)] != [(0.985 < (x / 50 + y / 175 + z / 12.5) / 8 < 1.015)]:
        triggered.add(98)
    # 原99、100被删除，从101开始顺延
    if [(min(x / 50, y / 175, z / 12.5) > 0.985)] != [(min(x / 50, y / 175, z / 12.5) > 9.85)]:
        triggered.add(99)
    if [(min(x / 50, y / 175, z / 12.5) > 0.985)] != [(min(x / 50, y / 175, z / 12.5) > 9.85)]:
        triggered.add(100)
    if [(max(x / 50, y / 175, z / 12.5) < 1.015)] != [(max(x / 50, y / 175, z / 12.5) < 15)]:
        triggered.add(101)
    if [(max(x / 50, y / 175, z / 12.5) < 1.015)] != [(max(x / 50, y / 175, z / 12.5) < 11.5)]:
        triggered.add(102)
    if [(abs(max(x, y / 4, z * 4) - min(x, y / 4, z * 4)) < 8)] != [(abs(max(x, y / 4, z * 4) - min(x, y / 4, z * 4)) < 18)]:
        triggered.add(103)
    if [(abs(max(x, y / 4, z * 4) - min(x, y / 4, z * 4)) < 8)] != [(abs(max(x, y / 4, z * 4) - min(x, y / 4, z * 14)) < 8)]:
        triggered.add(104)
    if [(278 < (x + y + z * 10) < 282)] != [(278 < (x + y + z * 10) < 482)]:
        triggered.add(105)
    if [(278 < (x + y + z * 10) < 282)] != [(278 < (x + y + z * 10) < 989)]:
        triggered.add(106)
    if [(24500 < (x * y * z) < 25500)] != [(24500 < (x * y * 6) < 25500)]:
        triggered.add(107)
    if [(24500 < (x * y * z) < 25500)] != [(24500 < (x * y * 2) < 25500)]:
        triggered.add(108)
    if [(abs((x + y + z * 10) / 3 - 93.3) < 1)] != [(abs((x + y + z * 10) / 3 - 93.3) < 11)]:
        triggered.add(109)
    if [(abs((x + y + z * 10) / 3 - 93.3) < 1)] != [(abs((x + y + z * 10) / 3 - 93.3) < 6)]:
        triggered.add(110)
    if [(49.5 < x < 50.5 and 174 < y < 176)] != [(49.5 < x < 50.5 and 174 < y < 196)]:
        triggered.add(111)
    if [(49.5 < x < 50.5 and 174 < y < 176)] != [(49.5 < x < 50.5 and 174 < y < 676)]:
        triggered.add(112)
    if [(12.45 < z < 12.55)] != [(12.45 < z < 128)]:
        triggered.add(113)
    if [(12.45 < z < 12.55)] != [(12.45 < z < 1255)]:
        triggered.add(114)
    if [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 16) < 16] != [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 4 * 16) < 16]:
        triggered.add(115)
    if [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 16) < 16] != [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 16) < 96]:
        triggered.add(116)
    if [(abs(x / (y / 4) / z - 0.091) < 0.003)] != [(abs(x / (y / 4) / z - 0.091) < 3)]:
        triggered.add(117)
    if [(abs(x / (y / 4) / z - 0.091) < 0.003)] != [(abs(x / (y / 4) / z - 0.091) < 1.3)]:
        triggered.add(118)
    if [(x < 45 or y < 160 or z < 11.5)] != [(x < 45 or y < 160 or z * 8 < 11.5)]:
        triggered.add(119)
    if [(x < 45 or y < 160 or z < 11.5)] != [(x < 45 or y * 8 < 160 or z < 11.5)]:
        triggered.add(120)
    if [(x > 55 or y > 190 or z > 13.5)] != [(x * 2 > 55 or y > 190 or z > 13.5)]:
        triggered.add(121)
    if [(x > 55 or y > 190 or z > 13.5)] != [(x > 55 or y * 8 > 190 or z > 13.5)]:
        triggered.add(122)
    if [(abs(x / (y / 4) - 1.14) > 0.15)] != [(abs(x / (y / 4) - 1.14) > 15)]:
        triggered.add(123)
    if [(abs(x / (y / 4) - 1.14) > 0.15)] != [(abs(x / (y / 4) - 1.14) > 10)]:
        triggered.add(124)
    if [(abs(x / z - 4.0) > 0.5)] != [(abs(x / z - 4.0) > 5)]:
        triggered.add(125)
    if [(abs(x / z - 4.0) > 0.5)] != [(abs(x / z - 4.0) > 21)]:
        triggered.add(126)
    if [(abs(y / z - 14) > 2)] != [(abs(y / z - 14) > 99)]:
        triggered.add(127)
    if [(abs(y / z - 14) > 2)] != [(abs(y / z - 14) > 7)]:
        triggered.add(128)
    if [((x + y + z * 10) < 270 or (x + y + z * 10) > 290)] != [((x + y + z * 10) < 270 or (x + y + z * 10) > 490)]:
        triggered.add(129)
    if [((x + y + z * 10) < 270 or (x + y + z * 10) > 290)] != [((x + y + z * 10) < 270 or (x + y + z * 10) > 890)]:
        triggered.add(130)
    if [((x * y * z) < 22000 or (x * y * z) > 28000)] != [((x * y * z) < 22000 or (x * y * z) > 98000)]:
        triggered.add(131)
    if [((x * y * z) < 22000 or (x * y * z) > 28000)] != [((x * y * z) < 22000 or (x * y * z) > 99000)]:
        triggered.add(132)
    if [(x < 43 and y < 160)] != [(x < 43 and y < 1680)]:
        triggered.add(133)
    if [(x < 43 and y < 160)] != [(x < 43 and y < 1600)]:
        triggered.add(134)
    if [(x > 57 and y > 190)] != [(x > 57 and y * 9 > 190)]:
        triggered.add(135)
    if [(x > 57 and y > 190)] != [(x > 57 and y * 8 > 190)]:
        triggered.add(136)
    if [(x < 43 and z < 11)] != [(x < 43 and z * 2 < 11)]:
        triggered.add(137)
    if [(x < 43 and z < 11)] != [(x < 43 and 8 < 11)]:
        triggered.add(138)
    if [(x > 57 and z > 14)] != [(x > 57 and z * 8 > 14)]:
        triggered.add(139)
    if [(x > 57 and z > 14)] != [(x > 57 and 59 > 14)]:
        triggered.add(140)
    if [(y < 160 and z < 11)] != [(y < 160 and z * 8 < 11)]:
        triggered.add(141)
    if [(y < 160 and z < 11)] != [(y < 160 and 9 < 11)]:
        triggered.add(142)
    if [(y > 190 and z > 14)] != [(y > 190 and z * 44 > 14)]:
        triggered.add(143)
    if [(y > 190 and z > 14)] != [(y > 190 and 23 > 14)]:
        triggered.add(144)
    if [(x < 38 or x > 62)] != [(x < 38 or x * 78 > 62)]:
        triggered.add(145)
    if [(x < 38 or x > 62)] != [(x < 38 or x > 162)]:
        triggered.add(146)
    if [(y < 145 or y > 205)] != [(y < 145 or y > 150)]:
        triggered.add(147)
    if [(y < 145 or y > 205)] != [(y < 115 or y > 205)]:
        triggered.add(148)
    if [(z < 9.5 or z > 15.5)] != [(z < 9.5 or z > 155)]:
        triggered.add(149)
    if [(z < 9.5 or z > 15.5)] != [(z < 9.5 or z > 95.5)]:
        triggered.add(150)
    if [(x < 35 and y < 155 and z < 10.5)] != [(x < 35 and y < 155 and z < 105)]:
        triggered.add(151)
    if [(x < 35 and y < 155 and z < 10.5)] != [(x < 35 and y < 155 and z * 8 < 10.5)]:
        triggered.add(152)
    if [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 16) > 100] != [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 16) > 900]:
        triggered.add(153)
    if [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 16) > 100] != [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 89) > 100]:
        triggered.add(154)
    if [(abs((x * y * z) / 25000 - 1) > 0.2)] != [(abs((x * y * z) / 25000 - 1) > 2)]:
        triggered.add(155)
    if [(abs((x * y * z) / 25000 - 1) > 0.2)] != [(abs((x * y * z) / 25000 - 1) > 7)]:
        triggered.add(156)

    return triggered

# === 执行函数绑定（使用 section10） ===
execute_Tr = section10_comprehensive_hybrid_control

# === 状态处理辅助函数 ===
def clip_state(state):
    return np.clip(state, EXPERIMENT_CONFIG['MIN_VALUES'], EXPERIMENT_CONFIG['MAX_VALUES'])

def denormalize_state(normalized_state):
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return normalized_state * (max_vals - min_vals) / 2 + (min_vals + max_vals) / 2

def coverage_similarity(triggered, target_path):
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0
    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)

def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']
    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']
    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']
    return reward

# === PPO Actor ===
class PPOActor(nn.Module):
    def __init__(self, state_dim=3, action_dim=3, hidden_dim=256):
        super(PPOActor, self).__init__()
        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)
        self.mean_head = nn.Linear(hidden_dim, action_dim)
        self.log_std_head = nn.Linear(hidden_dim, action_dim)
        self.action_scale = 10.0

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))
        x = F.relu(self.fc3(x))
        mean = torch.tanh(self.mean_head(x)) * self.action_scale
        log_std = torch.clamp(self.log_std_head(x), -20, 2)
        return mean, log_std

    def get_action_and_log_prob(self, state):
        mean, log_std = self.forward(state)
        std = torch.exp(log_std)
        dist = torch.distributions.Normal(mean, std)
        action = dist.sample()
        log_prob = dist.log_prob(action).sum(dim=-1)
        return action, log_prob

    def get_log_prob(self, state, action):
        mean, log_std = self.forward(state)
        std = torch.exp(log_std)
        dist = torch.distributions.Normal(mean, std)
        log_prob = dist.log_prob(action).sum(dim=-1)
        return log_prob

# === PPO Critic ===
class PPOCritic(nn.Module):
    def __init__(self, state_dim=3, hidden_dim=256):
        super(PPOCritic, self).__init__()
        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)
        self.value_head = nn.Linear(hidden_dim, 1)

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))
        x = F.relu(self.fc3(x))
        value = self.value_head(x)
        return value.squeeze(-1)

# === PPO Buffer ===
class PPOBuffer:
    def __init__(self):
        self.states = []
        self.actions = []
        self.rewards = []
        self.values = []
        self.log_probs = []
        self.dones = []
        self.path_indices = []
        self.similarities = []

    def store(self, state, action, reward, value, log_prob, done, path_idx, similarity):
        self.states.append(state)
        self.actions.append(action)
        self.rewards.append(reward)
        self.values.append(value)
        self.log_probs.append(log_prob)
        self.dones.append(done)
        self.path_indices.append(path_idx)
        self.similarities.append(similarity)

    def compute_advantages(self):
        rewards = np.array(self.rewards)
        values = np.array(self.values)
        dones = np.array(self.dones)
        advantages = np.zeros_like(rewards)
        returns = np.zeros_like(rewards)
        last_advantage = 0

        for t in reversed(range(len(rewards))):
            if t == len(rewards) - 1:
                next_value = 0
            else:
                next_value = values[t + 1]
            delta = rewards[t] + EXPERIMENT_CONFIG['GAMMA'] * next_value * (1 - dones[t]) - values[t]
            advantages[t] = delta + EXPERIMENT_CONFIG['GAMMA'] * EXPERIMENT_CONFIG['GAE_LAMBDA'] * (1 - dones[t]) * last_advantage
            last_advantage = advantages[t]

        returns = advantages + values
        advantages = (advantages - np.mean(advantages)) / (np.std(advantages) + 1e-8)
        return advantages, returns

    def get_batch_iterator(self, advantages, returns):
        indices = np.arange(len(self.states))
        np.random.shuffle(indices)
        batch_size = EXPERIMENT_CONFIG['PPO_BATCH_SIZE']
        for start in range(0, len(self.states), batch_size):
            end = min(start + batch_size, len(self.states))
            batch_indices = indices[start:end]
            yield {
                'states': torch.FloatTensor(np.array([self.states[i] for i in batch_indices])).to(device),
                'actions': torch.FloatTensor(np.array([self.actions[i] for i in batch_indices])).to(device),
                'old_log_probs': torch.FloatTensor(np.array([self.log_probs[i] for i in batch_indices])).to(device),
                'advantages': torch.FloatTensor(advantages[batch_indices]).to(device),
                'returns': torch.FloatTensor(returns[batch_indices]).to(device)
            }

    def get_top_k_per_path(self, num_paths, k=20):
        path_samples = {i: [] for i in range(num_paths)}
        for idx in range(len(self.states)):
            path_idx = self.path_indices[idx]
            similarity = self.similarities[idx]
            state = self.states[idx]
            path_samples[path_idx].append((idx, similarity, state))

        top_k_results = {}
        for path_idx in range(num_paths):
            samples = path_samples[path_idx]
            if len(samples) == 0:
                top_k_results[path_idx] = []
                continue
            samples.sort(key=lambda x: x[1], reverse=True)
            top_k = samples[:k]
            top_k_results[path_idx] = []
            for sample in top_k:
                normalized_state = sample[2]
                original_state = denormalize_state(normalized_state)
                # 保留浮点数，直接使用
                triggered = execute_Tr(*original_state)
                top_k_results[path_idx].append({
                    'state': original_state,
                    'similarity': sample[1],
                    'triggered': triggered
                })
        return top_k_results

    def clear(self):
        self.states.clear()
        self.actions.clear()
        self.rewards.clear()
        self.values.clear()
        self.log_probs.clear()
        self.dones.clear()
        self.path_indices.clear()
        self.similarities.clear()

    def __len__(self):
        return len(self.states)

# === PPO Agent ===
class PPOAgent:
    def __init__(self, state_dim=3, action_dim=3):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.actor = PPOActor(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.actor_optimizer = optim.Adam(self.actor.parameters(), lr=EXPERIMENT_CONFIG['ACTOR_LR'])
        self.critic = PPOCritic(state_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_optimizer = optim.Adam(self.critic.parameters(), lr=EXPERIMENT_CONFIG['CRITIC_LR'])
        self.buffer = PPOBuffer()
        self.update_count = 0

    def get_action(self, state):
        min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
        max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
        normalized_state = (state - (min_vals + max_vals) / 2) / ((max_vals - min_vals) / 2)
        state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)
        with torch.no_grad():
            action, log_prob = self.actor.get_action_and_log_prob(state_tensor)
            value = self.critic(state_tensor)
        action = action.cpu().numpy()[0]
        log_prob = log_prob.cpu().item()
        value = value.cpu().item()
        return action, log_prob, value

    def store_experience(self, state, action, reward, value, log_prob, done, path_idx, similarity):
        min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
        max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
        normalized_state = (state - (min_vals + max_vals) / 2) / ((max_vals - min_vals) / 2)
        self.buffer.store(normalized_state, action, reward, value, log_prob, done, path_idx, similarity)

    def update(self):
        if len(self.buffer) == 0:
            return
        advantages, returns = self.buffer.compute_advantages()
        for epoch in range(EXPERIMENT_CONFIG['PPO_EPOCHS']):
            for batch in self.buffer.get_batch_iterator(advantages, returns):
                new_log_probs = self.actor.get_log_prob(batch['states'], batch['actions'])
                ratio = torch.exp(new_log_probs - batch['old_log_probs'])
                surr1 = ratio * batch['advantages']
                surr2 = torch.clamp(ratio, 1 - EXPERIMENT_CONFIG['CLIP_EPSILON'], 1 + EXPERIMENT_CONFIG['CLIP_EPSILON']) * batch['advantages']
                actor_loss = -torch.min(surr1, surr2).mean()
                self.actor_optimizer.zero_grad()
                actor_loss.backward()
                torch.nn.utils.clip_grad_norm_(self.actor.parameters(), 0.5)
                self.actor_optimizer.step()

                new_values = self.critic(batch['states'])
                critic_loss = F.mse_loss(new_values, batch['returns'])
                self.critic_optimizer.zero_grad()
                critic_loss.backward()
                torch.nn.utils.clip_grad_norm_(self.critic.parameters(), 0.5)
                self.critic_optimizer.step()

        self.update_count += 1
        self.buffer.clear()
        if self.update_count % 2 == 0:
            print(f"  -> PPO completed (Run {self.update_count})")

# === Metric 计算函数 ===
def calculate_run_performance(run_idx, ppo_results, training_time, total_steps, update_count, agent):
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)
    total_reward = 0
    all_similarities = []
    total_samples = 0
    all_rewards = []

    for path_idx in range(num_paths):
        samples = ppo_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']
            total_reward += reward
            all_rewards.append(reward)
            all_similarities.append(similarity)
            total_samples += 1

    average_reward = total_reward / total_samples if total_samples > 0 else 0
    convergence = np.mean(all_similarities) if all_similarities else 0
    environment_adaptability = 1 / (np.std(all_similarities) + 1e-8) if len(all_similarities) > 1 else 0
    generalization_ability = convergence
    computational_efficiency = total_steps / training_time if training_time > 0 else 0
    policy_update_frequency = update_count / training_time if training_time > 0 else 0
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0

    return {
        'Run': run_idx + 1,
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }

# === Excel 导出函数 ===
def export_to_excel(all_ppo_results, all_performance_data, target_paths, output_path="PPO_20_run.xlsx"):
    print("\n导出Excel...")
    all_ppo_summary_data = []
    all_ppo_detailed_data = []

    for run_idx, (ppo_results, performance_data) in enumerate(zip(all_ppo_results, all_performance_data)):
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = ppo_results[path_idx]
            if len(samples) == 0:
                all_ppo_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue
            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = 'Yes' if perfect_count > 0 else 'No'
            all_ppo_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })
            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']
                all_ppo_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Sample ID': sample_idx + 1,
                    'X(湿度)': round(state[0], 1),
                    'Y(扭矩)': round(state[1], 1),
                    'Z(电流)': round(state[2], 2),
                    'Similarity': round(similarity, 4),
                    'Perfect Coverage': 'Yes' if abs(similarity - 1.0) < 0.001 else 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path))),
                    'Triggered Rules': ', '.join(map(str, sorted(triggered))),
                    'Hit Rule Count': len(target_path.intersection(triggered)),
                    'Target Rule Count': len(target_path)
                })

    dqn_summary_df = pd.DataFrame(all_ppo_summary_data)
    dqn_detailed_df = pd.DataFrame(all_ppo_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        dqn_summary_df.to_excel(writer, sheet_name='PPOPath', index=False)
        dqn_detailed_df.to_excel(writer, sheet_name='PPODetailed Sample Data', index=False)
        selected_columns = [
            'Run', 'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df[selected_columns].to_excel(writer, sheet_name='Metric', index=False)

        workbook = writer.book
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        for sheet_name in ['PPOPath', 'PPODetailed Sample Data', 'Metric']:
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

    print(f"Excel已保存: {output_path}")

# === 训练流程 ===
def train_ppo_workflow():
    print("=" * 80)
    print("PPO 训练 - 综合混合控制（section10）")
    print("状态范围: x(湿度) 1~65, y(扭矩) 100~205, z(电流) 1~20")
    print("相似度 = 命中规则数 / 目标路径规则数")
    print("=" * 80)

    agent = PPOAgent()
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n生成初始样本: {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']} / 路径")
    path_samples = {}
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']

    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            # 所有维度均为连续浮点数
            state = np.array([
                np.random.uniform(min_vals[0], max_vals[0]),
                np.random.uniform(min_vals[1], max_vals[1]),
                np.random.uniform(min_vals[2], max_vals[2])
            ], dtype=np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  路径 {path_idx + 1}/{num_paths}: {len(samples)} 个初始样本")

    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size

    print(f"\n批次大小: {batch_size}, 单样本迭代步数: {EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(f"总批次数: {num_batches} /路径 × {num_paths} 路径 = {num_batches * num_paths}")
    print("-" * 80)

    global_buffer = PPOBuffer()

    for batch_idx in range(num_batches):
        print(f"\n批次 {batch_idx + 1}/{num_batches}")

        for path_idx in range(num_paths):
            target_path = target_paths[path_idx]
            batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]

            batch_rewards = []
            batch_similarities = []

            for initial_state in batch_samples:
                state = initial_state.copy()
                episode_reward = 0
                final_similarity = 0

                for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                    action, log_prob, value = agent.get_action(state)
                    next_state = state + action
                    next_state = clip_state(next_state)

                    triggered = execute_Tr(*next_state)
                    reward = unified_reward_function(triggered, target_path)
                    similarity = coverage_similarity(triggered, target_path)
                    done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)

                    agent.store_experience(state, action, reward, value, log_prob, done, path_idx, similarity)
                    normalized_state = (state - (min_vals + max_vals) / 2) / ((max_vals - min_vals) / 2)
                    global_buffer.store(
                        normalized_state,
                        action, reward, value, log_prob, done, path_idx, similarity
                    )

                    state = next_state
                    episode_reward += reward
                    final_similarity = similarity
                    total_steps += 1

                batch_rewards.append(episode_reward)
                batch_similarities.append(final_similarity)

            avg_reward = np.mean(batch_rewards)
            avg_similarity = np.mean(batch_similarities)
            print(f"  路径 {path_idx + 1}: 平均奖励={avg_reward:.2f}, 平均相似度={avg_similarity:.4f}")

        print(f"\n  执行 PPO 策略更新...")
        agent.update()
        print(f"  全局缓冲区累计样本数: {len(global_buffer)}")

    training_time = time.time() - start_time

    print("\n" + "=" * 80)
    print(f"单次运行训练完成! 总耗时: {training_time:.2f} 秒, 总交互步数: {total_steps}")
    print(f"全局缓冲区大小: {len(global_buffer)}")
    print(f"PPO 策略更新次数: {agent.update_count}")
    print("=" * 80)

    print(f"\n提取每条路径的 Top-{EXPERIMENT_CONFIG['TOP_K_SAMPLES']} 高相似度样本...")
    ppo_top_k_results = global_buffer.get_top_k_per_path(num_paths, EXPERIMENT_CONFIG['TOP_K_SAMPLES'])

    return agent, ppo_top_k_results, training_time, total_steps, agent.update_count

# === 主程序 ===
def main():
    print("\n" + "=" * 80)
    print("PPO 多轮实验 - 综合混合控制（section10）")
    print("参数范围: 湿度 1~65, 扭矩 100~205, 电流 1~20")
    print("=" * 80)

    all_ppo_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'='*80}")
        print(f"开始第 {run_idx + 1}/{EXPERIMENT_CONFIG['NUM_RUNS']} 次独立运行")
        print(f"{'='*80}")

        ppo_agent, ppo_results, training_time, total_steps, update_count = train_ppo_workflow()
        performance_data = calculate_run_performance(
            run_idx, ppo_results, training_time, total_steps, update_count, ppo_agent
        )

        all_ppo_results.append(ppo_results)
        all_performance_data.append(performance_data)

        print(f"\n第 {run_idx + 1} 次运行完成!")
        print(f"  总奖励: {performance_data['Total Reward']}")
        print(f"  平均奖励: {performance_data['Average Reward']}")
        print(f"  平均相似度: {performance_data['Average Similarity']}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"PPO_section10_20run_{timestamp}.xlsx"
    export_to_excel(all_ppo_results, all_performance_data, target_paths, output_path)

    print("\n" + "=" * 80)
    print(f"{EXPERIMENT_CONFIG['NUM_RUNS']} 次运行整体统计摘要")
    print("=" * 80)

    total_rewards = [p['Total Reward'] for p in all_performance_data]
    average_rewards = [p['Average Reward'] for p in all_performance_data]
    convergences = [p['Convergence'] for p in all_performance_data]
    environment_adaptabilities = [p['Environment Adaptability'] for p in all_performance_data]
    generalization_abilities = [p['Generalization Ability'] for p in all_performance_data]
    computational_efficiencies = [p['Computational Efficiency'] for p in all_performance_data]
    policy_update_frequencies = [p['Policy Update Frequency'] for p in all_performance_data]
    avg_similarities = [p['Average Similarity'] for p in all_performance_data]

    print(f"\nTotal Reward Statistics:")
    print(f"  Mean: {np.mean(total_rewards):.2f}")
    print(f"  Std: {np.std(total_rewards):.2f}")

    print(f"\nAverage Reward Statistics:")
    print(f"  Mean: {np.mean(average_rewards):.4f}")
    print(f"  Std: {np.std(average_rewards):.4f}")

    print(f"\nConvergence Statistics:")
    print(f"  Mean: {np.mean(convergences):.4f}")
    print(f"  Std: {np.std(convergences):.4f}")

    print(f"\nEnvironment Adaptability Statistics:")
    print(f"  Mean: {np.mean(environment_adaptabilities):.4f}")
    print(f"  Std: {np.std(environment_adaptabilities):.4f}")

    print(f"\nGeneralization Ability Statistics:")
    print(f"  Mean: {np.mean(generalization_abilities):.4f}")
    print(f"  Std: {np.std(generalization_abilities):.4f}")

    print(f"\nComputational Efficiency Statistics:")
    print(f"  Mean: {np.mean(computational_efficiencies):.2f}")
    print(f"  Std: {np.std(computational_efficiencies):.2f}")

    print(f"\nPolicy Update Frequency Statistics:")
    print(f"  Mean: {np.mean(policy_update_frequencies):.4f}")
    print(f"  Std: {np.std(policy_update_frequencies):.4f}")

    print("\n" + "=" * 80)
    print(f" {EXPERIMENT_CONFIG['NUM_RUNS']} 次运行全部完成!")
    print("=" * 80)


if __name__ == "__main__":
    main()