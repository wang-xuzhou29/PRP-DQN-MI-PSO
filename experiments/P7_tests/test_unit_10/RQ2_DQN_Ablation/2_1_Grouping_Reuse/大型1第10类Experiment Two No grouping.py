import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import math

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === 全局取值范围（修改为 X:1~65, Y:100~205, Z:1~20） ===
MIN_X = 1
MAX_X = 65
MIN_Y = 100
MAX_Y = 205
MIN_Z = 1
MAX_Z = 20

# === 归一化/反归一化 ===
def normalize_state(state):
    """将状态归一化到 [0, 1] 区间"""
    weather_norm = (state[0] - MIN_X) / (MAX_X - MIN_X)
    time_norm = (state[1] - MIN_Y) / (MAX_Y - MIN_Y)
    z_norm = (state[2] - MIN_Z) / (MAX_Z - MIN_Z)
    return (weather_norm, time_norm, z_norm)

def denormalize_state(state_norm):
    """将归一化状态还原"""
    weather = int(round(state_norm[0] * (MAX_X - MIN_X) + MIN_X))
    time_period = int(round(state_norm[1] * (MAX_Y - MIN_Y) + MIN_Y))
    z = int(round(state_norm[2] * (MAX_Z - MIN_Z) + MIN_Z))

    # 边界保护
    weather = np.clip(weather, MIN_X, MAX_X)
    time_period = np.clip(time_period, MIN_Y, MAX_Y)
    z = np.clip(z, MIN_Z, MAX_Z)

    return (weather, time_period, z)

def normalize_value(value, min_val, max_val):
    return (value - min_val) / (max_val - min_val)

def denormalize_value(value_norm, min_val, max_val):
    return int(round(value_norm * (max_val - min_val) + min_val))

# === 安全除法 ===
def safe_divide(numerator, denominator, default=0.0):
    if denominator == 0:
        return default
    return numerator / denominator

# === reward function ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10

    if target_path.issubset(triggered):
        reward += 1

    if prev_triggered is not None:
        prev_sim = jaccard_similarity(prev_triggered, target_path)
        improvement = sim - prev_sim
        reward += improvement * 5

    return reward

# ========== 规则触发函数（section10_comprehensive_hybrid_control） ==========
def section10_comprehensive_hybrid_control(x, y, z):
    """第10类: 综合混合控制 (湿度x, 扭矩y, 电流z) - 156个分支（删除99、100后顺延编号）"""
    triggered = set()  # 用于记录触发的条件编号，便于问题溯源

    # -------------------------- 1-40：综合系统基础检查（每组2个） --------------------------
    # 1-2：湿度基础参数正常
    if [(30 < x < 70)] != [
        (30 < x * 78 < 70)]:
        triggered.add(1)
    if [(30 < x < 70)] != [
        (30 < 122 < 70)]:
        triggered.add(2)

        # 3-4：扭矩基础参数正常
    if [(140 < y < 210)] != [
        (140 < 678 < 210)]:
        triggered.add(3)
    if [(140 < y < 210)] != [
        (140 < y * 88 < 210)]:
        triggered.add(4)

        # 5-6：电流基础参数正常
    if [(8 < z < 16)] != [
        (8 < z * 67 < 16)]:
        triggered.add(5)
    if [(8 < z < 16)] != [
        (8 < 566 < 16)]:
        triggered.add(6)

        # 7-8：湿度扭矩核心区间协调（45-55%RH & 160-190N·m）
    if [(45 < x < 55 and 160 < y < 190)] != [
        (45 < x < 55 and 160 < y * 8 < 190)]:
        triggered.add(7)
    if [(45 < x < 55 and 160 < y < 190)] != [
        (45 < x * 6 < 55 and 160 < y < 190)]:
        triggered.add(8)

        # 9-10：湿度电流匹配良好（42-58%RH & 11-14A）
    if [(42 < x < 58 and 11 < z < 14)] != [
        (42 < x < 58 and 11 < z * 78 < 14)]:
        triggered.add(9)
    if [(42 < x < 58 and 11 < z < 14)] != [
        (42 < x < 58 and 11 < z * 8 < 14)]:
        triggered.add(10)

        # 11-12：扭矩电流协调稳定（155-195N·m & 11-14A）
    if [(155 < y < 195 and 11 < z < 14)] != [
        (155 < y < 195 and 11 < z * 78 < 14)]:
        triggered.add(11)
    if [(155 < y < 195 and 11 < z < 14)] != [
        (155 < y * 66 < 195 and 11 < z < 14)]:
        triggered.add(12)

        # 13-14：湿度精确控制（48-52%RH）
    if [(48 < x < 52)] != [(48 < x * 78 < 52)]:
        triggered.add(13)
    if [(48 < x < 52)] != [(48 < 88 < 52)]:
        triggered.add(14)

        # 15-16：扭矩精确控制（173-177N·m）
    if [(173 < y < 177)] != [(173 < y * 7 < 177)]:
        triggered.add(15)
    if [(173 < y < 177)] != [(173 < 88 < 177)]:
        triggered.add(16)

        # 17-18：电流精确控制（12.2-12.8A）
    if [(12.2 < z < 12.8)] != [(12.2 < z < 128)]:
        triggered.add(17)
    if [(12.2 < z < 12.8)] != [(10 < z < 12.8)]:
        triggered.add(18)

        # 19-20：湿度稳定性良好（偏离目标50%RH < 2%）
    if [(abs(x - 50) < 2)] != [(abs(x - 50) < 12)]:
        triggered.add(19)
    if [(abs(x - 50) < 2)] != [(abs(x - 50) < 8)]:
        triggered.add(20)

        # 21-22：扭矩稳定性良好（偏离目标175N·m < 5N·m）
    if [(abs(y - 175) < 5)] != [(abs(y - 175) < 15)]:
        triggered.add(21)
    if [(abs(y - 175) < 5)] != [(abs(y - 175) < 51)]:
        triggered.add(22)

        # 23-24：电流稳定性良好（偏离目标12.5A < 0.3A）
    if [(abs(z - 12.5) < 0.3)] != [(abs(z - 12.5) < 3)]:
        triggered.add(23)
    if [(abs(z - 12.5) < 0.3)] != [(abs(z - 12.5) < 1)]:
        triggered.add(24)

        # 25-26：所有参数在安全范围（下限：42%RH、155N·m、11A）
    if [(x > 42 and y > 155 and z > 11)] != [(x > 42 and y > 155 and 22 > 11)]:
        triggered.add(25)
    if [(x > 42 and y > 155 and z > 11)] != [(x > 42 and y > 155 and z * 3 > 11)]:
        triggered.add(26)

        # 27-28：所有参数未超上限（上限：58%RH、195N·m、14A）
    if [(x < 58 and y < 195 and z < 14)] != [(x < 58 and y < 195 and 10 < 14)]:
        triggered.add(27)
    if [(x < 58 and y < 195 and z < 14)] != [(x < 58 and y < 195 and z * 78 < 14)]:
        triggered.add(28)

        # 29-30：综合系统指标正常（x+y+10z：260-300）
    if [(260 < x + y + z * 10 < 300)] != [
        (260 < x + y + z * 10 < 3100)]:
        triggered.add(29)
    if [(260 < x + y + z * 10 < 300)] != [
        (300 < x + y + z * 10 < 300)]:
        triggered.add(30)

        # 31-32：湿度扭矩乘积正常（8000-10000）
    if [(8000 < x * y < 10000)] != [(8000 < x * y < 70000)]:
        triggered.add(31)
    if [(8000 < x * y < 10000)] != [(8000 < x * y < 12000)]:
        triggered.add(32)

        # 33-34：湿度电流乘积正常（600-750）
    if [(600 < x * z < 750)] != [(600 < x * z < 7510)]:
        triggered.add(33)
    if [(600 < x * z < 750)] != [(600 < x * z < 1750)]:
        triggered.add(34)

        # 35-36：扭矩电流乘积正常（2100-2500）
    if [(2100 < y * z < 2500)] != [(210 < y * z < 2500)]:
        triggered.add(35)
    if [(2100 < y * z < 2500)] != [(2100 < 23 * z < 2500)]:
        triggered.add(36)

        # 37-38：平均系统参数正常（(x+y+10z)/3：86-100）
    if [(86 < (x + y + z * 10) / 3 < 100)] != [
        (86 < (x + y + z * 10) / 3 < 1020)]:
        triggered.add(37)
    if [(86 < (x + y + z * 10) / 3 < 100)] != [
        (86 < (x + y + z * 10) / 3 < 200)]:
        triggered.add(38)

        # 39-40：系统向量模长正常（sqrt(x²+y²+(10z)²) > 200）
    if [(math.sqrt(x ** 2 + y ** 2 + (z * 10) ** 2) > 200)] != [
        (math.sqrt(x ** 2 + y ** 2 + (z * 10) ** 2) > 2100)]:
        triggered.add(39)
    if [(math.sqrt(x ** 2 + y ** 2 + (z * 10) ** 2) > 200)] != [
        (math.sqrt(x ** 2 + y ** 2 + (z * 10) ** 2) > 500)]:
        triggered.add(40)

        # -------------------------- 41-80：跨域参数分析（每组2个） --------------------------
        # 41-42：湿度扭矩比理想（x/(y/4)：1.0-1.3）
    if [(1.0 < x / (y / 4) < 1.3)] != [(1.0 < x / (y / 4) < 13)]:
        triggered.add(41)
    if [(1.0 < x / (y / 4) < 1.3)] != [(1.0 < x / (y / 4) < 333)]:
        triggered.add(42)

        # 43-44：湿度电流比正常（x/z：3.5-4.5）
    if [(3.5 < x / z < 4.5)] != [(3.5 < x / z < 45)]:
        triggered.add(43)
    if [(3.5 < x / z < 4.5)] != [(3.5 < x / z < 24.5)]:
        triggered.add(44)

        # 45-46：扭矩电流比适当（y/z：12-16）
    if [(12 < y / z < 16)] != [(12 < y / z < 30)]:
        triggered.add(45)
    if [(12 < y / z < 16)] != [(12 < y / z < 76)]:
        triggered.add(46)

        # 47-48：湿度扭矩偏差和正常（(x-40)+(y-150)/5：15-25）
    if [(15 < (x - 40) + (y - 150) / 5 < 25)] != [
        (15 < (x - 40) + (y - 150) / 5 < 215)]:
        triggered.add(47)
    if [(15 < (x - 40) + (y - 150) / 5 < 25)] != [
        (15 < (x - 40) + (y - 150) / 5 < 125)]:
        triggered.add(48)

        # 49-50：湿度电流偏差和正常（(x-40)+3(z-10)：15-25）
    if [(15 < (x - 40) + (z - 10) * 3 < 25)] != [
        (15 < (x - 40) + (z - 10) * 5 < 25)]:
        triggered.add(49)
    if [(15 < (x - 40) + (z - 10) * 3 < 25)] != [
        (15 < (x - 40) + (z - 10) * 10 < 25)]:
        triggered.add(50)

        # 51-52：扭矩电流偏差和正常（(y-150)/5+(z-10)：8-12）
    if [(8 < (y - 150) / 5 + (z - 10) < 12)] != [
        (8 < (y - 150) / 5 + (z - 10) < 121)]:
        triggered.add(51)
    if [(8 < (y - 150) / 5 + (z - 10) < 12)] != [
        (8 < (y - 150) / 5 + (z - 10) < 112)]:
        triggered.add(52)

        # 53-54：湿度扭矩偏差关系平衡（|(x-50)-(y-175)/4| < 5）
    if [(abs((x - 50) - (y - 175) / 4) < 5)] != [
        (abs((x - 50) - (y - 175) / 4) < 115)]:
        triggered.add(53)
    if [(abs((x - 50) - (y - 175) / 4) < 5)] != [
        (abs((x - 50) - (y - 175) / 4) < 51)]:
        triggered.add(54)

        # 55-56：湿度电流偏差关系平衡（|(x-50)-3(z-12.5)| < 6）
    if [(abs((x - 50) - (z - 12.5) * 3) < 6)] != [
        (abs((x - 50) - (z - 12.5) * 3) < 16)]:
        triggered.add(55)
    if [(abs((x - 50) - (z - 12.5) * 3) < 6)] != [
        (abs((x - 50) - (z - 12.5) * 3) < 26)]:
        triggered.add(56)

        # 57-58：扭矩电流偏差关系平衡（|(y-175)/14-(z-12.5)| < 2）
    if [(abs((y - 175) / 14 - (z - 12.5)) < 2)] != [
        (abs((y - 175) / 14 - (z - 12.5)) < 21)]:
        triggered.add(57)
    if [(abs((y - 175) / 14 - (z - 12.5)) < 2)] != [
        (abs((y - 175) / 14 - (z - 12.5)) < 112)]:
        triggered.add(58)

        # 59-60：调整湿度扭矩比正常（x/(y/4+10)：0.9-1.2）
    if [(0.9 < x / (y / 4 + 10) < 1.2)] != [
        (0.9 < x / (y / 4 + 10) < 12)]:
        triggered.add(59)
    if [(0.9 < x / (y / 4 + 10) < 1.2)] != [
        (0.9 < x / (y / 4 + 10) < 112)]:
        triggered.add(60)

        # 61-62：调整扭矩电流比正常（y/(z+5)：9-13）
    if [(9 < y / (z + 5) < 13)] != [(9 < y / (z + 5) < 53)]:
        triggered.add(61)
    if [(9 < y / (z + 5) < 13)] != [(9 < y / (z + 5) < 131)]:
        triggered.add(62)

        # 63-64：调整电流湿度比正常（z/(x/10)：2.2-2.8）
    if [(2.2 < z / (x / 10) < 2.8)] != [(2.2 < z / (x / 10) < 28)]:
        triggered.add(63)
    if [(2.2 < z / (x / 10) < 2.8)] != [(2.2 < z / (x / 10) < 128)]:
        triggered.add(64)

        # 65-66：三元系统积正常（x*y*z：20000-30000）
    if [(20000 < (x * y * z) < 30000)] != [
        (20000 < (x * y * z) < 56000)]:
        triggered.add(65)
    if [(20000 < (x * y * z) < 30000)] != [
        (20000 < (x * y * z) < 88000)]:
        triggered.add(66)

        # 67-68：加权系统和正常（0.8x+0.1y+8z：160-180）
    if [(160 < x * 0.8 + y * 0.1 + z * 8 < 180)] != [
        (160 < x * 0.8 + y * 1 + z * 8 < 180)]:
        triggered.add(67)
    if [(160 < x * 0.8 + y * 0.1 + z * 8 < 180)] != [
        (160 < x * 0.8 + y * 0.1 + z * 10 < 180)]:
        triggered.add(68)

        # 69-70：加权几何平均正常（(x/50)^0.3*(y/175)^0.4*(z/12.5)^0.3 > 0.9）
    if [((x / 50) ** 0.3 * (y / 175) ** 0.4 * (z / 12.5) ** 0.3 > 0.9)] != [
        ((x / 50) ** 0.3 * (y / 175) ** 0.4 * (z / 12.5) ** 0.3 > 9)]:
        triggered.add(69)
    if [((x / 50) ** 0.3 * (y / 175) ** 0.4 * (z / 12.5) ** 0.3 > 0.9)] != [
        ((x / 50) ** 0.3 * (y / 175) ** 0.4 * (z / 12.5) ** 0.3 > 19)]:
        triggered.add(70)

        # 71-72：湿度扭矩偏差积平衡（(x-50)(y-175)/4：-40-40）
    if [((x - 50) * (y - 175) / 4 > -40 and (x - 50) * (y - 175) / 4 < 40)] != [
        ((x - 50) * (y - 175) / 4 > -40 and (x - 50) * (y - 175) / 4 < 140)]:
        triggered.add(71)
    if [((x - 50) * (y - 175) / 4 > -40 and (x - 50) * (y - 175) / 4 < 40)] != [
        ((x - 50) * (y - 175) / 4 > -40 and (x - 50) * (y - 175) / 4 < 640)]:
        triggered.add(72)

        # 73-74：湿度电流偏差积平衡（(x-50)(z-12.5)：-15-15）
    if [((x - 50) * (z - 12.5) > -15 and (x - 50) * (z - 12.5) < 15)] != [
        ((x - 50) * (z - 12.5) > -15 and (x - 50) * (z - 12.5) < 115)]:
        triggered.add(73)
    if [((x - 50) * (z - 12.5) > -15 and (x - 50) * (z - 12.5) < 15)] != [
        ((x - 50) * (z - 12.5) > -15 and (x - 50) * (z - 12.5) < 151)]:
        triggered.add(74)

        # 75-76：扭矩电流偏差积平衡（(y-175)(z-12.5)/14：-8-8）
    if [((y - 175) / 14 * (z - 12.5) > -8 and (y - 175) / 14 * (z - 12.5) < 8)] != [
        ((y - 175) / 14 * (z - 12.5) > -8 and (y - 175) / 14 * (z - 12.5) < 118)]:
        triggered.add(75)
    if [((y - 175) / 14 * (z - 12.5) > -8 and (y - 175) / 14 * (z - 12.5) < 8)] != [
        ((y - 175) / 14 * (z - 12.5) > -8 and (y - 175) / 14 * (z - 12.5) < 138)]:
        triggered.add(76)

        # 77-78：归一化和接近理想（|x/50+y/175+z/12.5-3| < 0.15）
    if [(abs(x / 50 + y / 175 + z / 12.5 - 3) < 0.15)] != [
        (abs(x / 50 + y / 175 + z / 12.5 - 3) < 15)]:
        triggered.add(77)
    if [(abs(x / 50 + y / 175 + z / 12.5 - 3) < 0.15)] != [
        (abs(x / 50 + y / 175 + z / 12.5 - 3) < 325)]:
        triggered.add(78)

        # 79-80：归一化积接近理想（|(x/50)(y/175)(z/12.5)-1| < 0.1）
    if [(abs((x / 50) * (y / 175) * (z / 12.5) - 1) < 0.1)] != [
        (abs((x / 50) * (y / 175) * (z / 12.5) - 1) < 1)]:
        triggered.add(79)
    if [(abs((x / 50) * (y / 175) * (z / 12.5) - 1) < 0.1)] != [
        (abs((x / 50) * (y / 175) * (z / 12.5) - 1) < 12)]:
        triggered.add(80)

        # -------------------------- 81-156：全系统优化协调+智能预测控制（编号顺延） --------------------------
        # 81-82：全参数系统优化（湿度49-51%RH、扭矩173-177N·m、电流12.3-12.7A）
    if [(49 < x < 51 and 173 < y < 177 and 12 < z < 17)] != [
        (49 < x < 51 and 173 < y < 177 and 3 < z < 17)]:
        triggered.add(81)
    if [(49 < x < 51 and 173 < y < 177 and 12 < z < 17)] != [
        (25 < x < 51 and 173 < y < 177 and 12 < z < 17)]:
        triggered.add(82)

        # 83-84：湿度扭矩比与电流协调（|x/(y/4)-1.14|<0.1 & |z-12.5|<0.2）
    if [(abs(x / (y / 4) - 1.14) < 0.1 and abs(z - 12.5) < 0.2)] != [
        (abs(x / (y / 4) - 1.14) < 0.1 and abs(z - 12.5) < 1)]:
        triggered.add(83)
    if [(abs(x / (y / 4) - 1.14) < 0.1 and abs(z - 12.5) < 0.2)] != [
        (abs(x / (y / 4) - 1.14) < 0.1 and abs(z - 12.5) < 2)]:
        triggered.add(84)

        # 85-86：湿度扭矩积与电流协调（x*y：8700-8800 & 电流12.4-12.6A）
    if [(8700 < x * y < 8800 and 12.4 < z < 12.6)] != [
        (8700 < x * y < 8800 and 12.4 < z < 126)]:
        triggered.add(85)
    if [(8700 < x * y < 8800 and 12.4 < z < 12.6)] != [
        (8700 < x * y < 8800 and 12.4 < z < 32.6)]:
        triggered.add(86)

        # 87-88：湿度扭矩平均与电流协调（(x+y/4)/2：46-48 & 电流12.3-12.7A）
    if [(46 < (x + y / 4) / 2 < 48 and 12.3 < z < 12.7)] != [
        (46 < (x + y / 4) / 2 < 48 and 12.3 < z < 127)]:
        triggered.add(87)
    if [(46 < (x + y / 4) / 2 < 48 and 12.3 < z < 12.7)] != [
        (46 < (x + y / 4) / 2 < 48 and 12.3 < z < 72.7)]:
        triggered.add(88)

        # 89-90：湿度扭矩差与电流协调（|x-y/4|<6 & 电流12.3-12.7A）
    if [(abs(x - y / 4) < 6 and 12.3 < z < 12.7)] != [
        (abs(x - y / 4) < 6 and 12.3 < z < 127)]:
        triggered.add(89)
    if [(abs(x - y / 4) < 6 and 12.3 < z < 12.7)] != [
        (abs(x - y / 4) < 6 and 12.3 < z < 327)]:
        triggered.add(90)

        # 91-92：湿度扭矩距离与电流优秀（sqrt((x-50)²+(y/4-43.75)²)<2 & |z-12.5|<0.2）
    if [(math.sqrt((x - 50) ** 2 + (y / 4 - 43.75) ** 2) < 2 and abs(z - 12.5) < 0.2)] != [
        (math.sqrt((x - 50) ** 2 + (y / 4 - 43.75) ** 2) < 2 and abs(z - 12.5) < 2)]:
        triggered.add(91)
    if [(math.sqrt((x - 50) ** 2 + (y / 4 - 43.75) ** 2) < 2 and abs(z - 12.5) < 0.2)] != [
        (math.sqrt((x - 50) ** 2 + (y / 4 - 43.75) ** 2) < 2 and abs(z - 12.5) < 22)]:
        triggered.add(92)

        # 93-94：湿度扭矩相对值协调（相对偏差<2%）
    if [(0.98 < x / 50 < 1.02 and 0.985 < y / 175 < 1.015)] != [
        (0.98 < x / 50 < 1.02 and 0.985 < y / 175 < 12)]:
        triggered.add(93)
    if [(0.98 < x / 50 < 1.02 and 0.985 < y / 175 < 1.015)] != [
        (0.98 < x / 50 < 1.02 and 0.985 < y / 175 < 10)]:
        triggered.add(94)

        # 95-96：电流相对值协调（相对偏差<1.6%）
    if [(0.984 < z / 12.5 < 1.016)] != [(0.984 < z / 12.5 < 10.16)]:
        triggered.add(95)
    if [(0.984 < z / 12.5 < 1.016)] != [(0.984 < z / 12.5 < 4.016)]:
        triggered.add(96)

        # 97-98：归一化平均协调（平均相对偏差<1.5%）
    if [(0.985 < (x / 50 + y / 175 + z / 12.5) / 3 < 1.015)] != [
        (0.985 < (x / 50 + y / 175 + z / 12.5) / 3 < 10.15)]:
        triggered.add(97)
    if [(0.985 < (x / 50 + y / 175 + z / 12.5) / 3 < 1.015)] != [
        (0.985 < (x / 50 + y / 175 + z / 12.5) / 8 < 1.015)]:
        triggered.add(98)

        # 99-100：最小相对值协调良好（原101-102，编号顺延）
    if [(min(x / 50, y / 175, z / 12.5) > 0.985)] != [
        (min(x / 50, y / 175, z / 12.5) > 9.85)]:
        triggered.add(99)
    if [(min(x / 50, y / 175, z / 12.5) > 0.985)] != [
        (min(x / 50, y / 175, z / 12.5) > 9.85)]:
        triggered.add(100)

        # 101-102：最大相对值协调良好（原103-104，编号顺延）
    if [(max(x / 50, y / 175, z / 12.5) < 1.015)] != [
        (max(x / 50, y / 175, z / 12.5) < 15)]:
        triggered.add(101)
    if [(max(x / 50, y / 175, z / 12.5) < 1.015)] != [
        (max(x / 50, y / 175, z / 12.5) < 11.5)]:
        triggered.add(102)

        # 103-104：标准化后范围协调（原105-106，编号顺延）
    if [(abs(max(x, y / 4, z * 4) - min(x, y / 4, z * 4)) < 8)] != [
        (abs(max(x, y / 4, z * 4) - min(x, y / 4, z * 4)) < 18)]:
        triggered.add(103)
    if [(abs(max(x, y / 4, z * 4) - min(x, y / 4, z * 4)) < 8)] != [
        (abs(max(x, y / 4, z * 4) - min(x, y / 4, z * 14)) < 8)]:
        triggered.add(104)

        # 105-106：综合参数协调优秀（原107-108，编号顺延）
    if [(278 < (x + y + z * 10) < 282)] != [
        (278 < (x + y + z * 10) < 482)]:
        triggered.add(105)
    if [(278 < (x + y + z * 10) < 282)] != [
        (278 < (x + y + z * 10) < 989)]:
        triggered.add(106)

        # 107-108：三元积协调优秀（原109-110，编号顺延）
    if [(24500 < (x * y * z) < 25500)] != [
        (24500 < (x * y * 6) < 25500)]:
        triggered.add(107)
    if [(24500 < (x * y * z) < 25500)] != [
        (24500 < (x * y * 2) < 25500)]:
        triggered.add(108)

        # 109-110：平均值协调优秀（原111-112，编号顺延）
    if [(abs((x + y + z * 10) / 3 - 93.3) < 1)] != [
        (abs((x + y + z * 10) / 3 - 93.3) < 11)]:
        triggered.add(109)
    if [(abs((x + y + z * 10) / 3 - 93.3) < 1)] != [
        (abs((x + y + z * 10) / 3 - 93.3) < 6)]:
        triggered.add(110)

        # 111-112：湿度扭矩超精密协调（原113-114，编号顺延）
    if [(49.5 < x < 50.5 and 174 < y < 176)] != [
        (49.5 < x < 50.5 and 174 < y < 196)]:
        triggered.add(111)
    if [(49.5 < x < 50.5 and 174 < y < 176)] != [
        (49.5 < x < 50.5 and 174 < y < 676)]:
        triggered.add(112)

        # 113-114：电流超精密协调（原115-116，编号顺延）
    if [(12.45 < z < 12.55)] != [(12.45 < z < 128)]:
        triggered.add(113)
    if [(12.45 < z < 12.55)] != [(12.45 < z < 1255)]:
        triggered.add(114)

        # 115-116：三维距离协调完美（原117-118，编号顺延）
    if [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 16) < 16] != [
        ((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 4 * 16) < 16]:
        triggered.add(115)
    if [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 16) < 16] != [
        ((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 16) < 96]:
        triggered.add(116)

        # 117-118：连续比例协调完美（原119-120，编号顺延）
    if [(abs(x / (y / 4) / z - 0.091) < 0.003)] != [
        (abs(x / (y / 4) / z - 0.091) < 3)]:
        triggered.add(117)
    if [(abs(x / (y / 4) / z - 0.091) < 0.003)] != [
        (abs(x / (y / 4) / z - 0.091) < 1.3)]:
        triggered.add(118)

        # 119-120：检测到参数偏低趋势（原121-122，编号顺延）
    if [(x < 45 or y < 160 or z < 11.5)] != [(x < 45 or y < 160 or z * 8 < 11.5)]:
        triggered.add(119)
    if [(x < 45 or y < 160 or z < 11.5)] != [(x < 45 or y * 8 < 160 or z < 11.5)]:
        triggered.add(120)

        # 121-122：检测到参数偏高趋势（原123-124，编号顺延）
    if [(x > 55 or y > 190 or z > 13.5)] != [(x * 2 > 55 or y > 190 or z > 13.5)]:
        triggered.add(121)
    if [(x > 55 or y > 190 or z > 13.5)] != [(x > 55 or y * 8 > 190 or z > 13.5)]:
        triggered.add(122)

        # 123-124：湿度扭矩比需要调整（原125-126，编号顺延）
    if [(abs(x / (y / 4) - 1.14) > 0.15)] != [(abs(x / (y / 4) - 1.14) > 15)]:
        triggered.add(123)
    if [(abs(x / (y / 4) - 1.14) > 0.15)] != [(abs(x / (y / 4) - 1.14) > 10)]:
        triggered.add(124)

        # 125-126：湿度电流比需要调整（原127-128，编号顺延）
    if [(abs(x / z - 4.0) > 0.5)] != [(abs(x / z - 4.0) > 5)]:
        triggered.add(125)
    if [(abs(x / z - 4.0) > 0.5)] != [(abs(x / z - 4.0) > 21)]:
        triggered.add(126)

        # 127-128：扭矩电流比需要调整（原129-130，编号顺延）
    if [(abs(y / z - 14) > 2)] != [(abs(y / z - 14) > 99)]:
        triggered.add(127)
    if [(abs(y / z - 14) > 2)] != [(abs(y / z - 14) > 7)]:
        triggered.add(128)

        # 129-130：综合指标需要调整（原131-132，编号顺延）
    if [((x + y + z * 10) < 270 or (x + y + z * 10) > 290)] != [
        ((x + y + z * 10) < 270 or (x + y + z * 10) > 490)]:
        triggered.add(129)
    if [((x + y + z * 10) < 270 or (x + y + z * 10) > 290)] != [
        ((x + y + z * 10) < 270 or (x + y + z * 10) > 890)]:
        triggered.add(130)

        # 131-132：三元积需要调整（原133-134，编号顺延）
    if [((x * y * z) < 22000 or (x * y * z) > 28000)] != [
        ((x * y * z) < 22000 or (x * y * z) > 98000)]:
        triggered.add(131)
    if [((x * y * z) < 22000 or (x * y * z) > 28000)] != [
        ((x * y * z) < 22000 or (x * y * z) > 99000)]:
        triggered.add(132)

        # 133-134：湿度扭矩同时偏低（原135-136，编号顺延）
    if [(x < 43 and y < 160)] != [(x < 43 and y < 1680)]:
        triggered.add(133)
    if [(x < 43 and y < 160)] != [(x < 43 and y < 1600)]:
        triggered.add(134)

        # 135-136：湿度扭矩同时偏高（原137-138，编号顺延）
    if [(x > 57 and y > 190)] != [(x > 57 and y * 9 > 190)]:
        triggered.add(135)
    if [(x > 57 and y > 190)] != [(x > 57 and y * 8 > 190)]:
        triggered.add(136)

        # 137-138：湿度电流同时偏低（原139-140，编号顺延）
    if [(x < 43 and z < 11)] != [(x < 43 and z * 2 < 11)]:
        triggered.add(137)
    if [(x < 43 and z < 11)] != [(x < 43 and 8 < 11)]:
        triggered.add(138)

        # 139-140：湿度电流同时偏高（原141-142，编号顺延）
    if [(x > 57 and z > 14)] != [(x > 57 and z * 8 > 14)]:
        triggered.add(139)
    if [(x > 57 and z > 14)] != [(x > 57 and 59 > 14)]:
        triggered.add(140)

        # 141-142：扭矩电流同时偏低（原143-144，编号顺延）
    if [(y < 160 and z < 11)] != [(y < 160 and z * 8 < 11)]:
        triggered.add(141)
    if [(y < 160 and z < 11)] != [(y < 160 and 9 < 11)]:
        triggered.add(142)

        # 143-144：扭矩电流同时偏高（原145-146，编号顺延）
    if [(y > 190 and z > 14)] != [(y > 190 and z * 44 > 14)]:
        triggered.add(143)
    if [(y > 190 and z > 14)] != [(y > 190 and 23 > 14)]:
        triggered.add(144)

        # 145-146：湿度在临界范围（原147-148，编号顺延）
    if [(x < 38 or x > 62)] != [(x < 38 or x * 78 > 62)]:
        triggered.add(145)
    if [(x < 38 or x > 62)] != [(x < 38 or x > 162)]:
        triggered.add(146)

        # 147-148：扭矩在临界范围（原149-150，编号顺延）
    if [(y < 145 or y > 205)] != [(y < 145 or y > 150)]:
        triggered.add(147)
    if [(y < 145 or y > 205)] != [(y < 115 or y > 205)]:
        triggered.add(148)

        # 149-150：电流在临界范围（原151-152，编号顺延）
    if [(z < 9.5 or z > 15.5)] != [(z < 9.5 or z > 155)]:
        triggered.add(149)
    if [(z < 9.5 or z > 15.5)] != [(z < 9.5 or z > 95.5)]:
        triggered.add(150)

        # 151-152：所有参数严重偏低（原153-154，编号顺延）
    if [(x < 35 and y < 155 and z < 10.5)] != [
        (x < 35 and y < 155 and z < 105)]:
        triggered.add(151)
    if [(x < 35 and y < 155 and z < 10.5)] != [
        (x < 35 and y < 155 and z * 8 < 10.5)]:
        triggered.add(152)

        # 153-154：三维偏离过大（原157-158，编号顺延）
    if [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 16) > 100] != [
        ((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 16) > 900]:
        triggered.add(153)
    if [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 16) > 100] != [
        ((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 89) > 100]:
        triggered.add(154)

        # 155-156：三元积偏离过大（原159-160，编号顺延）
    if [(abs((x * y * z) / 25000 - 1) > 0.2)] != [
        (abs((x * y * z) / 25000 - 1) > 2)]:
        triggered.add(155)
    if [(abs((x * y * z) / 25000 - 1) > 0.2)] != [
        (abs((x * y * z) / 25000 - 1) > 7)]:
        triggered.add(156)

    return triggered

# === 目标路径组（section10 编号 1~156） ===
targetPaths = [
    {1, 2, 3, 4, 5, 6, 11, 12, 17, 21, 22, 23, 24, 29, 31, 32, 33, 34, 36, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48, 53,
     54, 55, 56, 67, 68, 69, 70, 71, 72, 77, 78, 79, 80, 95, 96, 97, 99, 100, 101, 102, 103, 105, 106, 113, 114, 117,
     118, 123, 124, 125, 126, 129, 130, 135, 136, 139, 140, 146, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 17, 22, 23, 29, 31, 32, 33, 34, 36, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48, 53, 54, 55, 56, 57,
     58, 59, 60, 67, 68, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 101, 102, 105, 106, 113, 114, 117, 118, 123, 124,
     125, 126, 127, 128, 129, 130, 135, 136, 139, 140, 146, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 11, 12, 17, 19, 21, 22, 23, 24, 29, 31, 32, 33, 34, 36, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48,
     53, 54, 55, 56, 67, 68, 69, 70, 77, 78, 79, 80, 95, 96, 97, 99, 100, 101, 102, 103, 105, 106, 113, 114, 117, 118,
     123, 124, 125, 126, 129, 130, 135, 136, 139, 140, 145, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 11, 12, 15, 16, 17, 23, 24, 29, 31, 32, 33, 34, 36, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48, 53,
     54, 55, 56, 59, 60, 67, 68, 69, 70, 77, 78, 79, 80, 95, 96, 97, 99, 100, 101, 102, 105, 106, 113, 114, 117, 118,
     123, 124, 125, 126, 129, 130, 135, 136, 139, 140, 146, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 11, 12, 17, 22, 23, 24, 29, 31, 33, 34, 36, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48, 53, 54, 55,
     56, 67, 68, 69, 70, 71, 72, 77, 78, 79, 80, 95, 96, 97, 99, 100, 101, 102, 103, 105, 106, 113, 114, 117, 118, 123,
     124, 125, 126, 129, 130, 139, 140, 143, 144, 146, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 11, 12, 17, 22, 23, 24, 29, 31, 32, 33, 34, 35, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48, 53, 54,
     55, 56, 59, 60, 67, 68, 69, 70, 77, 78, 79, 80, 95, 96, 97, 101, 102, 105, 106, 113, 114, 117, 118, 120, 123, 124,
     125, 126, 129, 130, 135, 136, 139, 140, 142, 146, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 17, 22, 23, 29, 31, 33, 34, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48, 51, 52, 53, 54, 55, 56, 69,
     70, 71, 72, 73, 74, 77, 78, 79, 80, 95, 96, 97, 99, 100, 101, 102, 103, 105, 106, 108, 113, 114, 117, 118, 123,
     124, 125, 126, 129, 130, 139, 140, 143, 144, 146, 147, 153, 156},

    {1, 2, 5, 6, 9, 10, 17, 19, 20, 22, 23, 24, 28, 29, 35, 37, 38, 41, 42, 49, 53, 54, 55, 56, 57, 58, 63, 64, 66, 68,
     71, 72, 77, 78, 79, 80, 95, 96, 101, 102, 105, 106, 109, 113, 114, 117, 118, 120, 121, 122, 123, 124, 125, 126,
     127, 128, 129, 130, 131, 132, 142, 145, 148, 153, 155, 156},

    {1, 2, 3, 4, 5, 6, 17, 19, 22, 23, 29, 33, 34, 35, 37, 38, 39, 40, 41, 42, 53, 54, 55, 56, 57, 58, 59, 60, 67, 68,
     69, 70, 73, 74, 77, 78, 79, 80, 85, 86, 95, 96, 97, 101, 102, 105, 106, 113, 114, 117, 118, 120, 123, 124, 127,
     128, 129, 130, 135, 136, 139, 140, 142, 145, 148, 153, 156},

    {1, 2, 3, 4, 17, 19, 21, 22, 29, 31, 32, 33, 34, 37, 38, 39, 40, 47, 48, 51, 52, 53, 54, 55, 56, 57, 58, 63, 64, 69,
     70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 99, 100, 101, 102, 105, 106, 113, 114, 117, 118, 123, 124, 125, 126, 127,
     128, 129, 130, 135, 136, 145, 147, 149, 150, 153, 156},

    {1, 2, 3, 4, 5, 6, 18, 19, 22, 23, 25, 26, 29, 31, 36, 37, 38, 39, 40, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 55,
     56, 57, 58, 68, 69, 70, 71, 72, 77, 78, 79, 80, 84, 97, 101, 102, 103, 105, 106, 108, 117, 118, 119, 125, 126, 127,
     128, 129, 130, 139, 140, 143, 144, 145, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 22, 23, 25, 26, 29, 31, 35, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 53, 54, 56, 57, 58,
     61, 62, 68, 69, 70, 71, 72, 77, 78, 79, 80, 97, 101, 102, 105, 106, 108, 117, 118, 119, 123, 124, 125, 126, 127,
     128, 129, 130, 139, 140, 143, 144, 146, 147, 153, 156},

    {1, 2, 3, 4, 17, 19, 22, 29, 33, 34, 37, 38, 39, 40, 41, 42, 53, 54, 55, 56, 57, 58, 59, 60, 63, 64, 69, 70, 73, 74,
     77, 78, 79, 80, 87, 88, 95, 96, 97, 101, 102, 105, 106, 113, 114, 117, 118, 120, 123, 124, 125, 126, 127, 128, 129,
     130, 135, 136, 142, 145, 148, 149, 150, 153, 156},

    {1, 2, 3, 4, 5, 6, 18, 22, 23, 24, 29, 31, 36, 37, 38, 39, 40, 43, 44, 45, 46, 47, 48, 51, 52, 53, 54, 55, 56, 57,
     58, 67, 68, 69, 70, 71, 72, 77, 78, 79, 80, 83, 84, 97, 101, 102, 103, 105, 106, 108, 117, 118, 125, 126, 127, 128,
     129, 130, 139, 140, 143, 144, 145, 147, 153, 156},

    {1, 2, 5, 6, 17, 19, 20, 22, 23, 27, 29, 35, 37, 38, 39, 40, 41, 42, 49, 50, 53, 54, 55, 56, 57, 58, 63, 64, 66, 67,
     68, 69, 70, 71, 72, 77, 78, 79, 80, 95, 96, 101, 102, 105, 106, 113, 114, 117, 118, 120, 123, 124, 125, 126, 127,
     128, 129, 130, 131, 132, 142, 145, 148, 153, 156},

    {1, 2, 3, 4, 13, 14, 17, 22, 29, 31, 32, 33, 34, 37, 38, 39, 40, 51, 52, 53, 54, 55, 56, 57, 58, 63, 64, 69, 70, 75,
     76, 77, 78, 79, 80, 89, 90, 93, 94, 95, 96, 97, 99, 100, 101, 102, 105, 106, 112, 113, 114, 117, 118, 125, 126,
     127, 128, 129, 130, 145, 147, 149, 150, 153, 156},

    {1, 2, 3, 4, 7, 8, 13, 14, 17, 21, 22, 27, 29, 33, 34, 37, 38, 39, 40, 51, 52, 56, 57, 58, 63, 64, 69, 70, 77, 78,
     79, 80, 87, 88, 89, 90, 92, 93, 94, 95, 96, 97, 99, 100, 101, 102, 105, 106, 111, 112, 113, 114, 117, 118, 125,
     126, 127, 128, 129, 130, 145, 147, 149, 150, 156},

    {1, 2, 5, 6, 9, 10, 17, 19, 20, 23, 24, 28, 30, 35, 41, 42, 49, 53, 54, 55, 56, 57, 58, 63, 64, 66, 68, 71, 72, 77,
     78, 79, 80, 95, 96, 101, 102, 105, 106, 109, 110, 113, 114, 117, 118, 121, 122, 123, 124, 125, 126, 127, 128, 129,
     130, 131, 132, 142, 145, 148, 153, 155, 156},

    {1, 2, 3, 4, 5, 6, 17, 22, 23, 24, 28, 29, 35, 37, 38, 50, 53, 54, 56, 57, 58, 63, 64, 66, 68, 71, 72, 77, 78, 79,
     80, 89, 90, 95, 96, 101, 102, 105, 106, 109, 113, 114, 117, 118, 121, 122, 123, 124, 125, 126, 127, 128, 129, 130,
     131, 132, 138, 142, 148, 151, 153, 155, 156},

    {1, 2, 3, 4, 5, 6, 11, 12, 15, 16, 17, 19, 23, 24, 28, 29, 36, 37, 38, 39, 40, 49, 53, 54, 55, 56, 63, 64, 69, 70,
     79, 80, 82, 89, 90, 95, 96, 101, 102, 103, 105, 106, 113, 114, 116, 117, 118, 121, 122, 123, 124, 125, 126, 129,
     130, 131, 132, 133, 134, 138, 145, 147, 156},

    {1, 2, 3, 4, 5, 6, 17, 19, 20, 22, 23, 27, 29, 37, 38, 39, 40, 49, 50, 51, 52, 53, 54, 55, 56, 63, 64, 67, 68, 69,
     70, 77, 78, 79, 80, 87, 88, 89, 90, 95, 96, 97, 101, 102, 103, 105, 106, 113, 114, 116, 117, 118, 123, 124, 125,
     126, 129, 130, 143, 144, 145, 147, 154, 156},

    {1, 2, 3, 4, 7, 8, 19, 20, 21, 22, 25, 26, 28, 30, 35, 39, 40, 43, 44, 46, 53, 54, 55, 56, 57, 58, 61, 62, 65, 66,
     73, 74, 77, 78, 79, 80, 101, 102, 104, 105, 106, 109, 110, 117, 118, 119, 121, 122, 123, 124, 125, 126, 127, 129,
     130, 131, 132, 145, 147, 153, 155, 156},

    {1, 2, 3, 4, 5, 6, 19, 21, 22, 28, 29, 35, 37, 38, 39, 40, 43, 44, 45, 46, 53, 54, 57, 58, 61, 62, 66, 73, 74, 77,
     78, 79, 80, 101, 102, 103, 104, 105, 106, 117, 118, 121, 122, 123, 124, 125, 126, 127, 128, 129, 130, 131, 132,
     133, 134, 137, 145, 147, 153, 155, 156},

    {1, 2, 5, 6, 17, 23, 29, 33, 34, 35, 37, 38, 41, 42, 43, 44, 53, 54, 55, 56, 57, 58, 59, 60, 67, 68, 69, 70, 73, 74,
     79, 80, 95, 96, 98, 101, 102, 105, 106, 109, 113, 114, 117, 118, 120, 123, 124, 127, 128, 129, 130, 131, 132, 135,
     136, 139, 140, 142, 146, 153, 156},

    {1, 2, 3, 4, 5, 6, 22, 25, 26, 29, 31, 32, 35, 37, 38, 41, 42, 43, 44, 45, 46, 47, 48, 50, 53, 54, 56, 57, 58, 59,
     60, 69, 70, 79, 80, 101, 102, 105, 106, 109, 117, 118, 123, 124, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136,
     139, 140, 141, 146, 147, 153, 156},

    {1, 2, 17, 19, 27, 29, 35, 37, 38, 39, 40, 41, 42, 49, 50, 53, 54, 56, 57, 58, 63, 64, 66, 67, 68, 72, 77, 78, 79,
     80, 95, 96, 101, 102, 105, 106, 107, 109, 113, 114, 117, 118, 123, 124, 125, 126, 127, 129, 130, 131, 132, 138,
     142, 145, 149, 150, 153, 155, 156},

    {1, 2, 3, 4, 5, 6, 7, 8, 13, 14, 15, 16, 23, 25, 26, 28, 29, 35, 37, 38, 39, 40, 43, 44, 45, 46, 55, 56, 57, 58, 66,
     69, 70, 77, 78, 79, 80, 81, 92, 103, 104, 105, 106, 117, 118, 119, 121, 122, 125, 126, 127, 128, 129, 130, 131,
     132, 145, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 17, 19, 20, 21, 22, 23, 24, 28, 29, 36, 37, 38, 39, 40, 49, 50, 53, 54, 67,
     68, 69, 70, 87, 88, 95, 96, 98, 101, 102, 103, 105, 106, 113, 114, 115, 116, 117, 118, 121, 122, 129, 130, 145,
     147, 156},

    {1, 2, 3, 4, 5, 6, 7, 8, 13, 14, 17, 23, 27, 29, 36, 37, 38, 39, 40, 49, 50, 63, 64, 67, 68, 69, 70, 79, 80, 84, 85,
     86, 87, 88, 89, 90, 91, 92, 95, 96, 97, 101, 102, 103, 105, 106, 113, 114, 116, 117, 118, 129, 130, 145, 147, 154,
     156},

    {3, 4, 5, 6, 22, 28, 30, 35, 45, 46, 53, 54, 55, 56, 63, 64, 65, 66, 71, 72, 73, 74, 77, 78, 79, 80, 103, 109, 110,
     117, 118, 121, 122, 123, 124, 125, 126, 127, 128, 131, 132, 137, 141, 147, 152, 153, 155, 156},
]

# 将路径列表转换为集合列表
target_paths = [set(path) for path in targetPaths]
NUM_PATHS = len(target_paths)

def jaccard_similarity(set1, set2):
    """Compute Jaccard similarity"""
    if not set2:
        return 0.0

    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0

# === Path Similarity ===
def compute_path_similarity_matrix(paths):
    """Path Similarity"""
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix

def group_paths_by_similarity(paths, threshold_percentile=50):
    """SimilarityPath """
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.percentile(avg_sim_scores, threshold_percentile)

    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]

    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)

    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group

# === Sample generation ===
def compute_robustness(state, path, sample_size=9):
    """计算鲁棒性"""
    base = section10_comprehensive_hybrid_control(state[0], state[1], state[2])
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0

    # 邻域偏移量（z范围1~20，步长适当缩小）
    deltas = [
        (-1, -1, -3), (0, -1, 0), (1, -1, 3),
        (-1, 0, -3), (1, 0, 3),
        (-1, 1, -3), (0, 1, 0), (1, 1, 3),
        (0, 0, 0)
    ]

    for dw, dt, dz in deltas[:sample_size]:
        if dw == dt == dz == 0:
            continue

        neighbor_weather = int(np.clip(state[0] + dw, MIN_X, MAX_X))
        neighbor_time = int(np.clip(state[1] + dt, MIN_Y, MAX_Y))
        neighbor_z = int(np.clip(state[2] + dz, MIN_Z, MAX_Z))
        neighbor = (neighbor_weather, neighbor_time, neighbor_z)

        n_trig = section10_comprehensive_hybrid_control(neighbor[0], neighbor[1], neighbor[2])
        if not n_trig:
            continue

        rob += jaccard_similarity(n_trig, base)
        neighbors += 1

    return rob / neighbors if neighbors > 0 else 0.0

def generate_samples_for_all_paths(num_candidates=2000, top_k=200, run_id=1):
    """为所有路径生成样本（权重筛选）"""
    BEST_WEIGHTS = [0.55, 0.25, 0.2]

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_individual.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Individual Path {path_id} (Weighted Screening) - Run {run_id}\n")
            f.write("weather time_period z\tScore\tSimilarity\tRobustness\tLengthDiff\n")
            for s in samples:
                weather, time_period, z = s['state']
                f.write(
                    f"{weather} {time_period} {z}\t{s['score']:.4f}\t{s['similarity']:.4f}\t"
                    f"{s['robustness']:.4f}\t{s['length_diff']:.4f}\n"
                )

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_individual"

    for path_idx in range(len(target_paths)):
        path = target_paths[path_idx]
        candidate_samples = []
        attempts = 0
        max_attempts = num_candidates * 10

        while len(candidate_samples) < num_candidates and attempts < max_attempts:
            attempts += 1

            weather = np.random.randint(MIN_X, MAX_X + 1)
            time_period = np.random.randint(MIN_Y, MAX_Y + 1)
            z = np.random.randint(MIN_Z, MAX_Z + 1)
            state = (weather, time_period, z)

            triggered = section10_comprehensive_hybrid_control(weather, time_period, z)
            if not triggered:
                continue

            sim = jaccard_similarity(triggered, path)
            rob = compute_robustness(state, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))

            candidate_samples.append({
                'state': state,
                'similarity': sim,
                'robustness': rob,
                'length_diff': len_diff,
                'triggered': triggered
            })

        if candidate_samples:
            for sample in candidate_samples:
                score = (BEST_WEIGHTS[0] * sample['similarity'] +
                         BEST_WEIGHTS[1] * sample['robustness'] +
                         BEST_WEIGHTS[2] * sample['length_diff'])
                sample['score'] = score

            candidate_samples.sort(key=lambda x: x['score'], reverse=True)
            selected_samples = candidate_samples[:top_k]
            save_samples(path_id=path_idx + 1, samples=selected_samples, base_dir=base_dir)

# === Shared Experience Replay ===
class SharedExperienceReplay:
    def __init__(self, capacity=20000):
        self.capacity = capacity
        self.buffer = deque(maxlen=self.capacity)
        self.priorities = deque(maxlen=self.capacity)

    def append(self, experience):
        self.buffer.append(experience)
        self.priorities.append(experience[-1])

    def sample(self, batch_size, alpha=0.6):
        if len(self.buffer) < batch_size:
            return [], [], []

        priorities = np.array(self.priorities, dtype=np.float64)
        priorities = np.power(priorities, alpha)
        probabilities = priorities / np.sum(priorities)

        batch_indices = np.random.choice(len(self.buffer), batch_size, p=probabilities, replace=False)
        batch = [self.buffer[idx] for idx in batch_indices]

        return batch, batch_indices, probabilities[batch_indices]

    def update_priorities(self, batch_indices, td_errors):
        for idx, td_error in zip(batch_indices, td_errors):
            if idx < len(self.priorities):
                self.priorities[idx] = max(abs(td_error), 1e-6)

    def __len__(self):
        return len(self.buffer)

    def get_high_reward_samples(self, target_path, num_samples=20):
        if len(self.buffer) == 0:
            return []

        samples_with_scores = []
        seen_states = set()

        for experience in self.buffer:
            state_tensor = experience[0]
            state_norm = state_tensor.cpu().numpy().flatten()
            state_tuple = denormalize_state((state_norm[0], state_norm[1], state_norm[2]))

            if state_tuple in seen_states:
                continue
            seen_states.add(state_tuple)

            triggered = section10_comprehensive_hybrid_control(state_tuple[0], state_tuple[1], state_tuple[2])
            reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)

            samples_with_scores.append((state_tuple, reward, sim, triggered))

        samples_with_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_scores[:num_samples]

def load_path_data(file_path):
    path_data = []

    if not os.path.exists(file_path):
        print(f"文件不存在: {file_path}")
        return path_data

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
            for line in lines[2:]:
                parts = line.strip().split("\t")
                if parts:
                    values = parts[0].split()
                    if len(values) >= 3:
                        state = (int(values[0]), int(values[1]), int(values[2]))
                        path_data.append(state)
    except Exception as e:
        print(f"读取文件失败 {file_path}: {e}")

    return path_data

# === DQN ===
class DQN(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dims=[128, 64]):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, hidden_dims[0])
        self.fc2 = nn.Linear(hidden_dims[0], hidden_dims[1])
        self.fc3 = nn.Linear(hidden_dims[1], action_dim)
        self.dropout = nn.Dropout(0.1)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = self.dropout(x)
        x = torch.relu(self.fc2(x))
        return self.fc3(x)

# === DQN Agent with PER ===
class DQNAgentWithPER:
    def __init__(self, state_dim, action_dim, replay_buffer,
                 gamma=0.99, epsilon=1.0, epsilon_decay=0.995,
                 epsilon_min=0.1, learning_rate=0.001, alpha=0.6, beta=0.4):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer
        self.alpha = alpha
        self.beta = beta

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        """
        动作解码：30个动作，分别对应三个维度的不同步长
        - weather: +/-1, 0(x2)
        - time_period: +/-1, 0(x2)
        - z: +/-3, +/-2, +/-1, 0(x2)  适配 1~20 范围
        """
        delta_values_weather_time = [1, 0, 0, -1]
        delta_values_z = [3, 2, 1, 0, 0, -1, -2, -3]

        dim = action_idx // 10
        delta_idx = action_idx % 10

        if dim == 0:  # weather
            if delta_idx >= 4:
                delta_idx = 3
            return (delta_values_weather_time[delta_idx], 0, 0)
        elif dim == 1:  # time_period
            if delta_idx >= 4:
                delta_idx = 3
            return (0, delta_values_weather_time[delta_idx], 0)
        elif dim == 2:  # z
            if delta_idx >= 8:
                delta_idx = 7
            return (0, 0, delta_values_z[delta_idx])

    def act(self, state_norm, legal_actions=None):
        if legal_actions is None:
            legal_actions = list(range(self.action_dim))

        if not legal_actions:
            return None

        if random.random() < self.epsilon:
            return random.choice(legal_actions)

        state_tensor = torch.tensor(state_norm, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)[0]

        legal_q_values = q_values[legal_actions]
        best_legal_idx = torch.argmax(legal_q_values).item()
        return legal_actions[best_legal_idx]

    def get_legal_actions(self, state):
        legal_actions = []

        for action_idx in range(self.action_dim):
            dw, dt, dz = self.decode_action(action_idx)

            next_weather = state[0] + dw
            next_time = state[1] + dt
            next_z = state[2] + dz

            if (MIN_X <= next_weather <= MAX_X and
                    MIN_Y <= next_time <= MAX_Y and
                    MIN_Z <= next_z <= MAX_Z):
                legal_actions.append(action_idx)

        return legal_actions

    def store_transition(self, state, action, reward, next_state, done):
        state_norm = normalize_state(state)
        next_state_norm = normalize_state(next_state)

        state_tensor = torch.tensor(state_norm, dtype=torch.float32).unsqueeze(0).to(device)
        next_state_tensor = torch.tensor(next_state_norm, dtype=torch.float32).unsqueeze(0).to(device)

        with torch.no_grad():
            q_values = self.model(state_tensor)
            next_q_values = self.target_model(next_state_tensor)
            max_next_q = next_q_values.max(1)[0]
            target_q = reward + (self.gamma * max_next_q * (1 - done))
            td_error = abs(q_values[0][action].item() - target_q.item())

        self.replay_buffer.append((state_tensor, action, reward, next_state_tensor, done, td_error))
        return td_error

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return 0.0

        batch, batch_indices, probabilities = self.replay_buffer.sample(batch_size, alpha=self.alpha)

        if not batch:
            return 0.0

        states, actions, rewards, next_states, dones, _ = zip(*batch)

        weights = (len(self.replay_buffer) * probabilities) ** (-self.beta)
        weights = weights / weights.max()
        weights = torch.tensor(weights, dtype=torch.float32).to(device)

        states = torch.cat(states).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.cat(next_states).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q = self.target_model(next_states).max(1)[0].detach()
        target_q = rewards + (self.gamma * next_max_q * (1 - dones))

        td_errors = current_q - target_q
        weighted_loss = (td_errors.pow(2) * weights).mean()

        self.optimizer.zero_grad()
        weighted_loss.backward()
        torch.nn.utils.clip_grad_norm_(self.model.parameters(), 1.0)
        self.optimizer.step()

        new_priorities = abs(td_errors.detach().cpu().numpy())
        self.replay_buffer.update_priorities(batch_indices, new_priorities)

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

        return weighted_loss.item()

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())

# === 训练函数 ===
def generate_and_train_for_individual_paths(path_documents, repeats=5, batch_size=32, run_id=1):
    state_dim = 3
    action_dim = 30

    shared_replay_buffer = SharedExperienceReplay(capacity=20000)
    agent = DQNAgentWithPER(state_dim, action_dim, shared_replay_buffer)

    total_cumulative_reward = 0
    path_rewards = {}

    print(f"\n=== 第 {run_id}/20 次训练开始 ===")
    start_time = time.time()

    SAMPLES_PER_BATCH = 50
    NUM_BATCHES = 4
    STEPS_PER_SAMPLE = 3

    for path_idx in range(len(target_paths)):
        path_id = path_idx + 1
        print(f"\n{'=' * 60}")
        print(f"路径 {path_id}/{NUM_PATHS}")
        print(f"{'=' * 60}")

        file_path = os.path.join(path_documents, f"path{path_id}_individual.txt")
        if not os.path.exists(file_path):
            print(f"  文件不存在: {file_path}")
            continue

        path_data = load_path_data(file_path)
        if not path_data:
            print(f"  路径 {path_id} 无有效数据")
            continue

        target_path = target_paths[path_idx]

        if path_idx not in path_rewards:
            path_rewards[path_idx] = 0

        for repeat_idx in range(repeats):
            print(f"\n  重复 {repeat_idx + 1}/{repeats}")

            for batch_idx in range(NUM_BATCHES):
                batch_start = batch_idx * SAMPLES_PER_BATCH
                batch_end = min(batch_start + SAMPLES_PER_BATCH, len(path_data))

                print(f"     批次 {batch_idx + 1}/{NUM_BATCHES} (样本 {batch_start}-{batch_end})")

                for sample_idx in range(batch_start, batch_end):
                    state = path_data[sample_idx]
                    prev_state = None
                    prev_triggered = None

                    for step in range(STEPS_PER_SAMPLE):
                        legal_actions = agent.get_legal_actions(state)

                        if not legal_actions:
                            break

                        state_norm = normalize_state(state)
                        action = agent.act(state_norm, legal_actions)
                        if action is None:
                            break

                        dw, dt, dz = agent.decode_action(action)

                        next_state = (
                            int(np.clip(state[0] + dw, MIN_X, MAX_X)),
                            int(np.clip(state[1] + dt, MIN_Y, MAX_Y)),
                            int(np.clip(state[2] + dz, MIN_Z, MAX_Z))
                        )

                        triggered = section10_comprehensive_hybrid_control(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered,
                                                prev_triggered, prev_state)
                        done = (step == STEPS_PER_SAMPLE - 1)

                        agent.store_transition(state, action, reward, next_state, done)

                        prev_state = state
                        prev_triggered = triggered
                        state = next_state

                        total_cumulative_reward += reward
                        path_rewards[path_idx] += reward

                if len(agent.replay_buffer) >= batch_size:
                    loss = agent.train(batch_size)
                    print(f"         批次 {batch_idx + 1} 完成，损失: {loss:.4f}")

                if (batch_idx + 1) % 2 == 0:
                    agent.update_target_model()
                    print(f"         目标网络已更新 (批次 {batch_idx + 1})")

        print(f"\n路径 {path_id} 完成，累计奖励: {path_rewards[path_idx]:.2f}")
        print(f"共享经验池大小: {len(shared_replay_buffer)}")

    training_time = time.time() - start_time
    print(f"\n=== 第 {run_id}/20 次训练完成，用时: {training_time:.2f} 秒 ===")

    return agent, shared_replay_buffer, total_cumulative_reward, path_rewards, training_time

# === Excel 报告生成 ===
def generate_excel_report(all_runs, similar_group, isolated_group, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    sim_paths = [i+1 for i in similar_group]
    iso_paths = [i+1 for i in isolated_group]

    wb = Workbook()
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    h_color = "4472C4"
    sim_color = "E2EFDA"
    iso_color = "FCE4D6"
    s_color = "FFF2CC"

    # Sheet1: 路径相似度
    ws1 = wb.active
    ws1.title = "路径相似度"
    headers1 = ['Path ID', '分组'] + [f'Run {i}' for i in range(1, 21)] + ['平均', '最高', '最低', '标准差']
    for c, h in enumerate(headers1, 1):
        cell = ws1.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=h_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for pid in range(1, NUM_PATHS + 1):
        row = pid + 1
        if pid in sim_paths:
            gtype, gcolor = "高相似组", sim_color
        elif pid in iso_paths:
            gtype, gcolor = "低相似组", iso_color
        else:
            gtype, gcolor = "未分组", "FFFFFF"

        ws1.cell(row, 1, f"Path {pid}").font = Font(bold=True)
        ws1.cell(row, 2, gtype)
        for c in [1, 2]:
            ws1.cell(row, c).fill = PatternFill("solid", fgColor=gcolor)
            ws1.cell(row, c).alignment = Alignment(horizontal="center")
            ws1.cell(row, c).border = border

        sims = []
        for ri, run in enumerate(all_runs):
            s = run['path_sims'].get(pid, {}).get('avg', 0.0)
            sims.append(s)
            cell = ws1.cell(row, 3 + ri, round(s, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        stats = [np.mean(sims), np.max(sims), np.min(sims), np.std(sims)]
        for i, v in enumerate(stats):
            cell = ws1.cell(row, 23 + i, round(v, 4))
            cell.number_format = '0.0000'
            cell.fill = PatternFill("solid", fgColor=s_color)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    # Sheet2: 分组统计
    ws2 = wb.create_sheet("分组统计")
    headers2 = ['分组', '包含路径'] + [f'Run {i}' for i in range(1, 21)] + ['平均相似度', '标准差']
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=h_color)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    def write_group_row(row, name, paths, color):
        ws2.cell(row, 1, name).font = Font(bold=True)
        ws2.cell(row, 2, ','.join(map(str, paths)))
        for c in [1, 2]:
            ws2.cell(row, c).fill = PatternFill("solid", fgColor=color)
            ws2.cell(row, c).alignment = Alignment(horizontal="center")
            ws2.cell(row, c).border = border
        vals = []
        for ri, run in enumerate(all_runs):
            v = np.mean([run['path_sims'].get(p, {}).get('avg', 0.0) for p in paths])
            vals.append(v)
            cell = ws2.cell(row, 3 + ri, round(v, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        cell = ws2.cell(row, 23, round(np.mean(vals), 4))
        cell.number_format = '0.0000'
        cell.fill = PatternFill("solid", fgColor=s_color)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        cell = ws2.cell(row, 24, round(np.std(vals), 4))
        cell.number_format = '0.0000'
        cell.fill = PatternFill("solid", fgColor=s_color)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    write_group_row(2, "高相似组", sim_paths, sim_color)
    write_group_row(3, "低相似组", iso_paths, iso_color)

    # Sheet3: 轮次汇总
    ws3 = wb.create_sheet("轮次汇总")
    headers3 = ['轮次', '耗时(秒)', '总体平均相似度', '最高相似度', '最低相似度', '高相似组平均', '低相似组平均', '回放池容量']
    for c, h in enumerate(headers3, 1):
        cell = ws3.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=h_color)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for ri, run in enumerate(all_runs, 1):
        row = ri + 1
        high_avg = np.mean([run['path_sims'].get(p, {}).get('avg', 0.0) for p in sim_paths])
        low_avg = np.mean([run['path_sims'].get(p, {}).get('avg', 0.0) for p in iso_paths]) if iso_paths else 0.0
        vals = [
            f"Run {ri}",
            round(run['time'], 2),
            round(run['overall_avg'], 4),
            round(run['max_sim'], 4),
            round(run['min_sim'], 4),
            round(high_avg, 4),
            round(low_avg, 4),
            20000
        ]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row, c, v)
            if c == 1:
                cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    # Sheet4: Top样本
    ws4 = wb.create_sheet("Top样本详情")
    headers4 = ['轮次', '路径', '序号', 'X', 'Y', 'Z', '相似度', '触发规则']
    for c, h in enumerate(headers4, 1):
        cell = ws4.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=h_color)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    r_idx = 2
    for ri, run in enumerate(all_runs, 1):
        for pid in range(1, NUM_PATHS + 1):
            samples = run['samples'].get(pid, [])
            pcolor = sim_color if pid in sim_paths else (iso_color if pid in iso_paths else "FFFFFF")
            for si, (st, _, sim, trig) in enumerate(samples, 1):
                x, y, z = st
                ws4.cell(r_idx, 1, f"Run {ri}").fill = PatternFill("solid", fgColor=pcolor)
                ws4.cell(r_idx, 2, f"Path {pid}").fill = PatternFill("solid", fgColor=pcolor)
                ws4.cell(r_idx, 3, si)
                ws4.cell(r_idx, 4, x)
                ws4.cell(r_idx, 5, y)
                ws4.cell(r_idx, 6, z)
                ws4.cell(r_idx, 7, round(sim, 4)).number_format = '0.0000'
                ws4.cell(r_idx, 8, ','.join(map(str, sorted(trig))))
                for c in range(1, 9):
                    ws4.cell(r_idx, c).alignment = Alignment(horizontal="center")
                    ws4.cell(r_idx, c).border = border
                r_idx += 1

    out_path = os.path.join(out_dir, "20轮训练汇总报告.xlsx")
    wb.save(out_path)
    print(f"Excel 报告已保存: {out_path}")

# === 20 轮总入口 ===
def run_20_times_training():
    model_dir = r"D:\Experiment\CNN\DQNNEW\saved_models_new_vars"
    sample_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_individual"
    report_dir = r"D:\Experiment\CNN\ComparisonExperiment2\excel_reports_new_vars"

    os.makedirs(model_dir, exist_ok=True)
    os.makedirs(sample_dir, exist_ok=True)
    os.makedirs(report_dir, exist_ok=True)

    similar_group, isolated_group = group_paths_by_similarity(target_paths)

    print("=" * 60)
    print("DQN 路径覆盖优化 - 20轮完整训练")
    print(f"参数范围: X [{MIN_X},{MAX_X}] | Y [{MIN_Y},{MAX_Y}] | Z [{MIN_Z},{MAX_Z}]")
    print(f"路径总数: {NUM_PATHS}")
    print(f"高相似组: {[i+1 for i in similar_group]}")
    print(f"低相似组: {[i+1 for i in isolated_group]}")
    print("=" * 60)

    all_runs = []

    for run_id in range(1, 21):
        print(f"\n===== 第 {run_id}/20 轮 =====")
        print("[1/2] 生成初始样本...")
        generate_samples_for_all_paths(num_candidates=2000, top_k=200, run_id=run_id)

        print("[2/2] 开始训练...")
        agent, buffer, total_r, path_r, elapsed = generate_and_train_for_individual_paths(
            sample_dir, repeats=5, batch_size=32, run_id=run_id
        )

        # 保存模型
        model_path = os.path.join(model_dir, f"model_run_{run_id}.pth")
        torch.save({
            'model_state_dict': agent.model.state_dict(),
            'optimizer_state_dict': agent.optimizer.state_dict(),
            'epsilon': agent.epsilon,
            'ranges': {'x': [MIN_X, MAX_X], 'y': [MIN_Y, MAX_Y], 'z': [MIN_Z, MAX_Z]}
        }, model_path)

        # 收集统计
        run_data = {'path_sims': {}, 'samples': {}, 'time': elapsed}
        all_sims = []
        for pi in range(NUM_PATHS):
            pid = pi + 1
            samples = buffer.get_high_reward_samples(target_paths[pi], 20)
            if samples:
                sims = [s[2] for s in samples]
                run_data['path_sims'][pid] = {
                    'avg': np.mean(sims),
                    'max': np.max(sims),
                    'min': np.min(sims)
                }
                run_data['samples'][pid] = samples
                all_sims.extend(sims)
            else:
                run_data['path_sims'][pid] = {'avg': 0.0, 'max': 0.0, 'min': 0.0}
                run_data['samples'][pid] = []

        run_data['overall_avg'] = np.mean(all_sims) if all_sims else 0.0
        run_data['max_sim'] = np.max(all_sims) if all_sims else 0.0
        run_data['min_sim'] = np.min(all_sims) if all_sims else 0.0
        all_runs.append(run_data)

        print(f"本轮完成 | 耗时: {elapsed:.1f}s | 总体平均相似度: {run_data['overall_avg']:.4f}")

    print("\n生成最终 Excel 报告...")
    generate_excel_report(all_runs, similar_group, isolated_group, report_dir)
    print("\n全部 20 轮训练完成！")

if __name__ == "__main__":
    run_20_times_training()