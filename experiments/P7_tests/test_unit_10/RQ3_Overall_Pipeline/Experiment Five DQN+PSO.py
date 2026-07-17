import os
import sys
import random
import time
from collections import deque
from datetime import datetime
from typing import Any, Dict, List, Optional, Set, Tuple

import numpy as np
import torch
import torch.nn as nn
import torch.optim as optim
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ========================================
# ========== experiment configuration parameters ==========
# ========================================
NUM_RUNS = 20

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === three-dimensional range settings ===
# Keep the current DQN state range used by the second script. To use a 0-500 range, modify this section only.
LIGHT_MIN = 1
LIGHT_MAX = 65
MOISTURE_MIN = 100
MOISTURE_MAX = 205
TEMP_MIN = 1
TEMP_MAX = 20
BOUNDS = {
    "light": (LIGHT_MIN, LIGHT_MAX),
    "moisture": (MOISTURE_MIN, MOISTURE_MAX),
    "temp": (TEMP_MIN, TEMP_MAX),
}

# === standard PSO parameters ===
PSO_W = 0.7
PSO_C1 = 1.5
PSO_C2 = 1.5
PSO_VMAX_RATIO = 0.2


def normalize_state(state):
    """Normalize the state to the [0, 1] range."""
    return np.array([
        (state[0] - BOUNDS["light"][0]) / (BOUNDS["light"][1] - BOUNDS["light"][0]),
        (state[1] - BOUNDS["moisture"][0]) / (BOUNDS["moisture"][1] - BOUNDS["moisture"][0]),
        (state[2] - BOUNDS["temp"][0]) / (BOUNDS["temp"][1] - BOUNDS["temp"][0]),
    ], dtype=np.float32)


def denormalize_state(normalized_state):
    """Restore the normalized state to the original range."""
    return np.array([
        normalized_state[0] * (BOUNDS["light"][1] - BOUNDS["light"][0]) + BOUNDS["light"][0],
        normalized_state[1] * (BOUNDS["moisture"][1] - BOUNDS["moisture"][0]) + BOUNDS["moisture"][0],
        normalized_state[2] * (BOUNDS["temp"][1] - BOUNDS["temp"][0]) + BOUNDS["temp"][0],
    ], dtype=float)


def generate_random_state():
    """Generate a random state within each dimension range."""
    light = np.random.randint(BOUNDS["light"][0], BOUNDS["light"][1] + 1)
    moisture = np.random.randint(BOUNDS["moisture"][0], BOUNDS["moisture"][1] + 1)
    temp = np.random.randint(BOUNDS["temp"][0], BOUNDS["temp"][1] + 1)
    return np.array([light, moisture, temp], dtype=float)


def clip_state(state):
    """Clip the state to each dimension boundary."""
    return np.array([
        np.clip(state[0], BOUNDS["light"][0], BOUNDS["light"][1]),
        np.clip(state[1], BOUNDS["moisture"][0], BOUNDS["moisture"][1]),
        np.clip(state[2], BOUNDS["temp"][0], BOUNDS["temp"][1]),
    ], dtype=float)


def is_state_valid(state):
    """Check whether the state is within all valid dimension ranges."""
    return (
        BOUNDS["light"][0] <= state[0] <= BOUNDS["light"][1]
        and BOUNDS["moisture"][0] <= state[1] <= BOUNDS["moisture"][1]
        and BOUNDS["temp"][0] <= state[2] <= BOUNDS["temp"][1]
    )


def execute_Tr(x, y, z):
    """第10类: 综合混合控制 (湿度x, 扭矩y, 电流z) - 158个分支（删除99、100后顺延编号）"""
    triggered = set()  # 用于记录触发的条件编号，便于问题溯源

    # -------------------------- 1-40：综合系统基础检查（每组2个） --------------------------
    # 1-2：湿度基础参数正常
    if [(30 < x < 70)] != [(30 < x * 78 < 70)]:
        triggered.add(1)
    if [(30 < x < 70)] != [(30 < 122 < 70)]:
        triggered.add(2)

    # 3-4：扭矩基础参数正常
    if [(140 < y < 210)] != [(140 < 678 < 210)]:
        triggered.add(3)
    if [(140 < y < 210)] != [(140 < y * 88 < 210)]:
        triggered.add(4)

    # 5-6：电流基础参数正常
    if [(8 < z < 16)] != [(8 < z * 67 < 16)]:
        triggered.add(5)
    if [(8 < z < 16)] != [(8 < 566 < 16)]:
        triggered.add(6)

    # 7-8：湿度扭矩核心区间协调（45-55%RH & 160-190N·m）
    if [(45 < x < 55 and 160 < y < 190)] != [(45 < x < 55 and 160 < y * 8 < 190)]:
        triggered.add(7)
    if [(45 < x < 55 and 160 < y < 190)] != [(45 < x * 6 < 55 and 160 < y < 190)]:
        triggered.add(8)

    # 9-10：湿度电流匹配良好（42-58%RH & 11-14A）
    if [(42 < x < 58 and 11 < z < 14)] != [(42 < x < 58 and 11 < z * 78 < 14)]:
        triggered.add(9)
    if [(42 < x < 58 and 11 < z < 14)] != [(42 < x < 58 and 11 < z * 8 < 14)]:
        triggered.add(10)

    # 11-12：扭矩电流协调稳定（155-195N·m & 11-14A）
    if [(155 < y < 195 and 11 < z < 14)] != [(155 < y < 195 and 11 < z * 78 < 14)]:
        triggered.add(11)
    if [(155 < y < 195 and 11 < z < 14)] != [(155 < y * 66 < 195 and 11 < z < 14)]:
        triggered.add(12)

    # 13-14：湿度精确控制（48-52%RH）
    if [(48 < x < 52)] != [(48 < x * 78 < 52)]:
        triggered.add(13)
    if [(48 < x < 52)] != [(48 < 88 < 52)]:
        triggered.add(14)

    # 15-16：扭矩精确控制（173-177N·m）
    if [(173 < y < 177)] != [(173 < y * 7 < 177)]:
        triggered.add(15)
    if [(173 < y < 177)] != [(173 < 88 < 177)]:
        triggered.add(16)

    # 17-18：电流精确控制（12.2-12.8A）
    if [(12.2 < z < 12.8)] != [(12.2 < z < 128)]:
        triggered.add(17)
    if [(12.2 < z < 12.8)] != [(10 < z < 12.8)]:
        triggered.add(18)

    # 19-20：湿度稳定性良好（偏离目标50%RH < 2%）
    if [(abs(x - 50) < 2)] != [(abs(x - 50) < 12)]:
        triggered.add(19)
    if [(abs(x - 50) < 2)] != [(abs(x - 50) < 8)]:
        triggered.add(20)

    # 21-22：扭矩稳定性良好（偏离目标175N·m < 5N·m）
    if [(abs(y - 175) < 5)] != [(abs(y - 175) < 15)]:
        triggered.add(21)
    if [(abs(y - 175) < 5)] != [(abs(y - 175) < 51)]:
        triggered.add(22)

    # 23-24：电流稳定性良好（偏离目标12.5A < 0.3A）
    if [(abs(z - 12.5) < 0.3)] != [(abs(z - 12.5) < 3)]:
        triggered.add(23)
    if [(abs(z - 12.5) < 0.3)] != [(abs(z - 12.5) < 1)]:
        triggered.add(24)

    # 25-26：所有参数在安全范围（下限：42%RH、155N·m、11A）
    if [(x > 42 and y > 155 and z > 11)] != [(x > 42 and y > 155 and 22 > 11)]:
        triggered.add(25)
    if [(x > 42 and y > 155 and z > 11)] != [(x > 42 and y > 155 and z * 3 > 11)]:
        triggered.add(26)

    # 27-28：所有参数未超上限（上限：58%RH、195N·m、14A）
    if [(x < 58 and y < 195 and z < 14)] != [(x < 58 and y < 195 and 10 < 14)]:
        triggered.add(27)
    if [(x < 58 and y < 195 and z < 14)] != [(x < 58 and y < 195 and z * 78 < 14)]:
        triggered.add(28)

    # 29-30：综合系统指标正常（x+y+10z：260-300）
    if [(260 < x + y + z * 10 < 300)] != [(260 < x + y + z * 10 < 3100)]:
        triggered.add(29)
    if [(260 < x + y + z * 10 < 300)] != [(300 < x + y + z * 10 < 300)]:
        triggered.add(30)

    # 31-32：湿度扭矩乘积正常（8000-10000）
    if [(8000 < x * y < 10000)] != [(8000 < x * y < 70000)]:
        triggered.add(31)
    if [(8000 < x * y < 10000)] != [(8000 < x * y < 12000)]:
        triggered.add(32)

    # 33-34：湿度电流乘积正常（600-750）
    if [(600 < x * z < 750)] != [(600 < x * z < 7510)]:
        triggered.add(33)
    if [(600 < x * z < 750)] != [(600 < x * z < 1750)]:
        triggered.add(34)

    # 35-36：扭矩电流乘积正常（2100-2500）
    if [(2100 < y * z < 2500)] != [(210 < y * z < 2500)]:
        triggered.add(35)
    if [(2100 < y * z < 2500)] != [(2100 < 23 * z < 2500)]:
        triggered.add(36)

    # 37-38：平均系统参数正常（(x+y+10z)/3：86-100）
    if [(86 < (x + y + z * 10) / 3 < 100)] != [(86 < (x + y + z * 10) / 3 < 1020)]:
        triggered.add(37)
    if [(86 < (x + y + z * 10) / 3 < 100)] != [(86 < (x + y + z * 10) / 3 < 200)]:
        triggered.add(38)

    # 39-40：系统向量模长正常（sqrt(x²+y²+(10z)²) > 200）
    if [(math.sqrt(x ** 2 + y ** 2 + (z * 10) ** 2) > 200)] != [(math.sqrt(x ** 2 + y ** 2 + (z * 10) ** 2) > 2100)]:
        triggered.add(39)
    if [(math.sqrt(x ** 2 + y ** 2 + (z * 10) ** 2) > 200)] != [(math.sqrt(x ** 2 + y ** 2 + (z * 10) ** 2) > 500)]:
        triggered.add(40)

    # -------------------------- 41-80：跨域参数分析（每组2个） --------------------------
    # 41-42：湿度扭矩比理想（x/(y/4)：1.0-1.3）
    if [(1.0 < x / (y / 4) < 1.3)] != [(1.0 < x / (y / 4) < 13)]:
        triggered.add(41)
    if [(1.0 < x / (y / 4) < 1.3)] != [(1.0 < x / (y / 4) < 333)]:
        triggered.add(42)

    # 43-44：湿度电流比正常（x/z：3.5-4.5）
    if [(3.5 < x / z < 4.5)] != [(3.5 < x / z < 45)]:
        triggered.add(43)
    if [(3.5 < x / z < 4.5)] != [(3.5 < x / z < 24.5)]:
        triggered.add(44)

    # 45-46：扭矩电流比适当（y/z：12-16）
    if [(12 < y / z < 16)] != [(12 < y / z < 30)]:
        triggered.add(45)
    if [(12 < y / z < 16)] != [(12 < y / z < 76)]:
        triggered.add(46)

    # 47-48：湿度扭矩偏差和正常（(x-40)+(y-150)/5：15-25）
    if [(15 < (x - 40) + (y - 150) / 5 < 25)] != [(15 < (x - 40) + (y - 150) / 5 < 215)]:
        triggered.add(47)
    if [(15 < (x - 40) + (y - 150) / 5 < 25)] != [(15 < (x - 40) + (y - 150) / 5 < 125)]:
        triggered.add(48)

    # 49-50：湿度电流偏差和正常（(x-40)+3(z-10)：15-25）
    if [(15 < (x - 40) + (z - 10) * 3 < 25)] != [(15 < (x - 40) + (z - 10) * 5 < 25)]:
        triggered.add(49)
    if [(15 < (x - 40) + (z - 10) * 3 < 25)] != [(15 < (x - 40) + (z - 10) * 10 < 25)]:
        triggered.add(50)

    # 51-52：扭矩电流偏差和正常（(y-150)/5+(z-10)：8-12）
    if [(8 < (y - 150) / 5 + (z - 10) < 12)] != [(8 < (y - 150) / 5 + (z - 10) < 121)]:
        triggered.add(51)
    if [(8 < (y - 150) / 5 + (z - 10) < 12)] != [(8 < (y - 150) / 5 + (z - 10) < 112)]:
        triggered.add(52)

    # 53-54：湿度扭矩偏差关系平衡（|(x-50)-(y-175)/4| < 5）
    if [(abs((x - 50) - (y - 175) / 4) < 5)] != [(abs((x - 50) - (y - 175) / 4) < 115)]:
        triggered.add(53)
    if [(abs((x - 50) - (y - 175) / 4) < 5)] != [(abs((x - 50) - (y - 175) / 4) < 51)]:
        triggered.add(54)

    # 55-56：湿度电流偏差关系平衡（|(x-50)-3(z-12.5)| < 6）
    if [(abs((x - 50) - (z - 12.5) * 3) < 6)] != [(abs((x - 50) - (z - 12.5) * 3) < 16)]:
        triggered.add(55)
    if [(abs((x - 50) - (z - 12.5) * 3) < 6)] != [(abs((x - 50) - (z - 12.5) * 3) < 26)]:
        triggered.add(56)

    # 57-58：扭矩电流偏差关系平衡（|(y-175)/14-(z-12.5)| < 2）
    if [(abs((y - 175) / 14 - (z - 12.5)) < 2)] != [(abs((y - 175) / 14 - (z - 12.5)) < 21)]:
        triggered.add(57)
    if [(abs((y - 175) / 14 - (z - 12.5)) < 2)] != [(abs((y - 175) / 14 - (z - 12.5)) < 112)]:
        triggered.add(58)

    # 59-60：调整湿度扭矩比正常（x/(y/4+10)：0.9-1.2）
    if [(0.9 < x / (y / 4 + 10) < 1.2)] != [(0.9 < x / (y / 4 + 10) < 12)]:
        triggered.add(59)
    if [(0.9 < x / (y / 4 + 10) < 1.2)] != [(0.9 < x / (y / 4 + 10) < 112)]:
        triggered.add(60)

    # 61-62：调整扭矩电流比正常（y/(z+5)：9-13）
    if [(9 < y / (z + 5) < 13)] != [(9 < y / (z + 5) < 53)]:
        triggered.add(61)
    if [(9 < y / (z + 5) < 13)] != [(9 < y / (z + 5) < 131)]:
        triggered.add(62)

    # 63-64：调整电流湿度比正常（z/(x/10)：2.2-2.8）
    if [(2.2 < z / (x / 10) < 2.8)] != [(2.2 < z / (x / 10) < 28)]:
        triggered.add(63)
    if [(2.2 < z / (x / 10) < 2.8)] != [(2.2 < z / (x / 10) < 128)]:
        triggered.add(64)

    # 65-66：三元系统积正常（x*y*z：20000-30000）
    if [(20000 < (x * y * z) < 30000)] != [(20000 < (x * y * z) < 56000)]:
        triggered.add(65)
    if [(20000 < (x * y * z) < 30000)] != [(20000 < (x * y * z) < 88000)]:
        triggered.add(66)

    # 67-68：加权系统和正常（0.8x+0.1y+8z：160-180）
    if [(160 < x * 0.8 + y * 0.1 + z * 8 < 180)] != [(160 < x * 0.8 + y * 1 + z * 8 < 180)]:
        triggered.add(67)
    if [(160 < x * 0.8 + y * 0.1 + z * 8 < 180)] != [(160 < x * 0.8 + y * 0.1 + z * 10 < 180)]:
        triggered.add(68)

    # 69-70：加权几何平均正常（(x/50)^0.3*(y/175)^0.4*(z/12.5)^0.3 > 0.9）
    if [((x / 50) ** 0.3 * (y / 175) ** 0.4 * (z / 12.5) ** 0.3 > 0.9)] != [((x / 50) ** 0.3 * (y / 175) ** 0.4 * (z / 12.5) ** 0.3 > 9)]:
        triggered.add(69)
    if [((x / 50) ** 0.3 * (y / 175) ** 0.4 * (z / 12.5) ** 0.3 > 0.9)] != [((x / 50) ** 0.3 * (y / 175) ** 0.4 * (z / 12.5) ** 0.3 > 19)]:
        triggered.add(70)

    # 71-72：湿度扭矩偏差积平衡（(x-50)(y-175)/4：-40-40）
    if [((x - 50) * (y - 175) / 4 > -40 and (x - 50) * (y - 175) / 4 < 40)] != [((x - 50) * (y - 175) / 4 > -40 and (x - 50) * (y - 175) / 4 < 140)]:
        triggered.add(71)
    if [((x - 50) * (y - 175) / 4 > -40 and (x - 50) * (y - 175) / 4 < 40)] != [((x - 50) * (y - 175) / 4 > -40 and (x - 50) * (y - 175) / 4 < 640)]:
        triggered.add(72)

    # 73-74：湿度电流偏差积平衡（(x-50)(z-12.5)：-15-15）
    if [((x - 50) * (z - 12.5) > -15 and (x - 50) * (z - 12.5) < 15)] != [((x - 50) * (z - 12.5) > -15 and (x - 50) * (z - 12.5) < 115)]:
        triggered.add(73)
    if [((x - 50) * (z - 12.5) > -15 and (x - 50) * (z - 12.5) < 15)] != [((x - 50) * (z - 12.5) > -15 and (x - 50) * (z - 12.5) < 151)]:
        triggered.add(74)

    # 75-76：扭矩电流偏差积平衡（(y-175)(z-12.5)/14：-8-8）
    if [((y - 175) / 14 * (z - 12.5) > -8 and (y - 175) / 14 * (z - 12.5) < 8)] != [((y - 175) / 14 * (z - 12.5) > -8 and (y - 175) / 14 * (z - 12.5) < 118)]:
        triggered.add(75)
    if [((y - 175) / 14 * (z - 12.5) > -8 and (y - 175) / 14 * (z - 12.5) < 8)] != [((y - 175) / 14 * (z - 12.5) > -8 and (y - 175) / 14 * (z - 12.5) < 138)]:
        triggered.add(76)

    # 77-78：归一化和接近理想（|x/50+y/175+z/12.5-3| < 0.15）
    if [(abs(x / 50 + y / 175 + z / 12.5 - 3) < 0.15)] != [(abs(x / 50 + y / 175 + z / 12.5 - 3) < 15)]:
        triggered.add(77)
    if [(abs(x / 50 + y / 175 + z / 12.5 - 3) < 0.15)] != [(abs(x / 50 + y / 175 + z / 12.5 - 3) < 325)]:
        triggered.add(78)

    # 79-80：归一化积接近理想（|(x/50)(y/175)(z/12.5)-1| < 0.1）
    if [(abs((x / 50) * (y / 175) * (z / 12.5) - 1) < 0.1)] != [(abs((x / 50) * (y / 175) * (z / 12.5) - 1) < 1)]:
        triggered.add(79)
    if [(abs((x / 50) * (y / 175) * (z / 12.5) - 1) < 0.1)] != [(abs((x / 50) * (y / 175) * (z / 12.5) - 1) < 12)]:
        triggered.add(80)

    # -------------------------- 81-158：全系统优化协调+智能预测控制（编号顺延） --------------------------
    # 81-82：全参数系统优化（湿度49-51%RH、扭矩173-177N·m、电流12.3-12.7A）
    if [(49 < x < 51 and 173 < y < 177 and 12 < z < 17)] != [(49 < x < 51 and 173 < y < 177 and 3 < z < 17)]:
        triggered.add(81)
    if [(49 < x < 51 and 173 < y < 177 and 12 < z < 17)] != [(25 < x < 51 and 173 < y < 177 and 12 < z < 17)]:
        triggered.add(82)

    # 83-84：湿度扭矩比与电流协调（|x/(y/4)-1.14|<0.1 & |z-12.5|<0.2）
    if [(abs(x / (y / 4) - 1.14) < 0.1 and abs(z - 12.5) < 0.2)] != [(abs(x / (y / 4) - 1.14) < 0.1 and abs(z - 12.5) < 1)]:
        triggered.add(83)
    if [(abs(x / (y / 4) - 1.14) < 0.1 and abs(z - 12.5) < 0.2)] != [(abs(x / (y / 4) - 1.14) < 0.1 and abs(z - 12.5) < 2)]:
        triggered.add(84)

    # 85-86：湿度扭矩积与电流协调（x*y：8700-8800 & 电流12.4-12.6A）
    if [(8700 < x * y < 8800 and 12.4 < z < 12.6)] != [(8700 < x * y < 8800 and 12.4 < z < 126)]:
        triggered.add(85)
    if [(8700 < x * y < 8800 and 12.4 < z < 12.6)] != [(8700 < x * y < 8800 and 12.4 < z < 32.6)]:
        triggered.add(86)

    # 87-88：湿度扭矩平均与电流协调（(x+y/4)/2：46-48 & 电流12.3-12.7A）
    if [(46 < (x + y / 4) / 2 < 48 and 12.3 < z < 12.7)] != [(46 < (x + y / 4) / 2 < 48 and 12.3 < z < 127)]:
        triggered.add(87)
    if [(46 < (x + y / 4) / 2 < 48 and 12.3 < z < 12.7)] != [(46 < (x + y / 4) / 2 < 48 and 12.3 < z < 72.7)]:
        triggered.add(88)

    # 89-90：湿度扭矩差与电流协调（|x-y/4|<6 & 电流12.3-12.7A）
    if [(abs(x - y / 4) < 6 and 12.3 < z < 12.7)] != [(abs(x - y / 4) < 6 and 12.3 < z < 127)]:
        triggered.add(89)
    if [(abs(x - y / 4) < 6 and 12.3 < z < 12.7)] != [(abs(x - y / 4) < 6 and 12.3 < z < 327)]:
        triggered.add(90)

    # 91-92：湿度扭矩距离与电流优秀（sqrt((x-50)²+(y/4-43.75)²)<2 & |z-12.5|<0.2）
    if [(math.sqrt((x - 50) ** 2 + (y / 4 - 43.75) ** 2) < 2 and abs(z - 12.5) < 0.2)] != [(math.sqrt((x - 50) ** 2 + (y / 4 - 43.75) ** 2) < 2 and abs(z - 12.5) < 2)]:
        triggered.add(91)
    if [(math.sqrt((x - 50) ** 2 + (y / 4 - 43.75) ** 2) < 2 and abs(z - 12.5) < 0.2)] != [(math.sqrt((x - 50) ** 2 + (y / 4 - 43.75) ** 2) < 2 and abs(z - 12.5) < 22)]:
        triggered.add(92)

    # 93-94：湿度扭矩相对值协调（相对偏差<2%）
    if [(0.98 < x / 50 < 1.02 and 0.985 < y / 175 < 1.015)] != [(0.98 < x / 50 < 1.02 and 0.985 < y / 175 < 12)]:
        triggered.add(93)
    if [(0.98 < x / 50 < 1.02 and 0.985 < y / 175 < 1.015)] != [(0.98 < x / 50 < 1.02 and 0.985 < y / 175 < 10)]:
        triggered.add(94)

    # 95-96：电流相对值协调（相对偏差<1.6%）
    if [(0.984 < z / 12.5 < 1.016)] != [(0.984 < z / 12.5 < 10.16)]:
        triggered.add(95)
    if [(0.984 < z / 12.5 < 1.016)] != [(0.984 < z / 12.5 < 4.016)]:
        triggered.add(96)

    # 97-98：归一化平均协调（平均相对偏差<1.5%）
    if [(0.985 < (x / 50 + y / 175 + z / 12.5) / 3 < 1.015)] != [(0.985 < (x / 50 + y / 175 + z / 12.5) / 3 < 10.15)]:
        triggered.add(97)
    if [(0.985 < (x / 50 + y / 175 + z / 12.5) / 3 < 1.015)] != [(0.985 < (x / 50 + y / 175 + z / 12.5) / 8 < 1.015)]:
        triggered.add(98)

    # 99-100：最小相对值协调良好（原101-102，编号顺延）
    if [(min(x / 50, y / 175, z / 12.5) > 0.985)] != [(min(x / 50, y / 175, z / 12.5) > 9.85)]:
        triggered.add(99)
    if [(min(x / 50, y / 175, z / 12.5) > 0.985)] != [(min(x / 50, y / 175, z / 12.5) > 9.85)]:
        triggered.add(100)

    # 101-102：最大相对值协调良好（原103-104，编号顺延）
    if [(max(x / 50, y / 175, z / 12.5) < 1.015)] != [(max(x / 50, y / 175, z / 12.5) < 15)]:
        triggered.add(101)
    if [(max(x / 50, y / 175, z / 12.5) < 1.015)] != [(max(x / 50, y / 175, z / 12.5) < 11.5)]:
        triggered.add(102)

    # 103-104：标准化后范围协调（原105-106，编号顺延）
    if [(abs(max(x, y / 4, z * 4) - min(x, y / 4, z * 4)) < 8)] != [(abs(max(x, y / 4, z * 4) - min(x, y / 4, z * 4)) < 18)]:
        triggered.add(103)
    if [(abs(max(x, y / 4, z * 4) - min(x, y / 4, z * 4)) < 8)] != [(abs(max(x, y / 4, z * 4) - min(x, y / 4, z * 14)) < 8)]:
        triggered.add(104)

    # 105-106：综合参数协调优秀（原107-108，编号顺延）
    if [(278 < (x + y + z * 10) < 282)] != [(278 < (x + y + z * 10) < 482)]:
        triggered.add(105)
    if [(278 < (x + y + z * 10) < 282)] != [(278 < (x + y + z * 10) < 989)]:
        triggered.add(106)

    # 107-108：三元积协调优秀（原109-110，编号顺延）
    if [(24500 < (x * y * z) < 25500)] != [(24500 < (x * y * 6) < 25500)]:
        triggered.add(107)
    if [(24500 < (x * y * z) < 25500)] != [(24500 < (x * y * 2) < 25500)]:
        triggered.add(108)

    # 109-110：平均值协调优秀（原111-112，编号顺延）
    if [(abs((x + y + z * 10) / 3 - 93.3) < 1)] != [(abs((x + y + z * 10) / 3 - 93.3) < 11)]:
        triggered.add(109)
    if [(abs((x + y + z * 10) / 3 - 93.3) < 1)] != [(abs((x + y + z * 10) / 3 - 93.3) < 6)]:
        triggered.add(110)

    # 111-112：湿度扭矩超精密协调（原113-114，编号顺延）
    if [(49.5 < x < 50.5 and 174 < y < 176)] != [(49.5 < x < 50.5 and 174 < y < 196)]:
        triggered.add(111)
    if [(49.5 < x < 50.5 and 174 < y < 176)] != [(49.5 < x < 50.5 and 174 < y < 676)]:
        triggered.add(112)

    # 113-114：电流超精密协调（原115-116，编号顺延）
    if [(12.45 < z < 12.55)] != [(12.45 < z < 128)]:
        triggered.add(113)
    if [(12.45 < z < 12.55)] != [(12.45 < z < 1255)]:
        triggered.add(114)

    # 115-116：三维距离协调完美（原117-118，编号顺延）
    if [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 16) < 16] != [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 4 * 16) < 16]:
        triggered.add(115)
    if [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 16) < 16] != [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 16) < 96]:
        triggered.add(116)

    # 117-118：连续比例协调完美（原119-120，编号顺延）
    if [(abs(x / (y / 4) / z - 0.091) < 0.003)] != [(abs(x / (y / 4) / z - 0.091) < 3)]:
        triggered.add(117)
    if [(abs(x / (y / 4) / z - 0.091) < 0.003)] != [(abs(x / (y / 4) / z - 0.091) < 1.3)]:
        triggered.add(118)

    # 119-120：检测到参数偏低趋势（原121-122，编号顺延）
    if [(x < 45 or y < 160 or z < 11.5)] != [(x < 45 or y < 160 or z * 8 < 11.5)]:
        triggered.add(119)
    if [(x < 45 or y < 160 or z < 11.5)] != [(x < 45 or y * 8 < 160 or z < 11.5)]:
        triggered.add(120)

    # 121-122：检测到参数偏高趋势（原123-124，编号顺延）
    if [(x > 55 or y > 190 or z > 13.5)] != [(x * 2 > 55 or y > 190 or z > 13.5)]:
        triggered.add(121)
    if [(x > 55 or y > 190 or z > 13.5)] != [(x > 55 or y * 8 > 190 or z > 13.5)]:
        triggered.add(122)

    # 123-124：湿度扭矩比需要调整（原125-126，编号顺延）
    if [(abs(x / (y / 4) - 1.14) > 0.15)] != [(abs(x / (y / 4) - 1.14) > 15)]:
        triggered.add(123)
    if [(abs(x / (y / 4) - 1.14) > 0.15)] != [(abs(x / (y / 4) - 1.14) > 10)]:
        triggered.add(124)

    # 125-126：湿度电流比需要调整（原127-128，编号顺延）
    if [(abs(x / z - 4.0) > 0.5)] != [(abs(x / z - 4.0) > 5)]:
        triggered.add(125)
    if [(abs(x / z - 4.0) > 0.5)] != [(abs(x / z - 4.0) > 21)]:
        triggered.add(126)

    # 127-128：扭矩电流比需要调整（原129-130，编号顺延）
    if [(abs(y / z - 14) > 2)] != [(abs(y / z - 14) > 99)]:
        triggered.add(127)
    if [(abs(y / z - 14) > 2)] != [(abs(y / z - 14) > 7)]:
        triggered.add(128)

    # 129-130：综合指标需要调整（原131-132，编号顺延）
    if [((x + y + z * 10) < 270 or (x + y + z * 10) > 290)] != [((x + y + z * 10) < 270 or (x + y + z * 10) > 490)]:
        triggered.add(129)
    if [((x + y + z * 10) < 270 or (x + y + z * 10) > 290)] != [((x + y + z * 10) < 270 or (x + y + z * 10) > 890)]:
        triggered.add(130)

    # 131-132：三元积需要调整（原133-134，编号顺延）
    if [((x * y * z) < 22000 or (x * y * z) > 28000)] != [((x * y * z) < 22000 or (x * y * z) > 98000)]:
        triggered.add(131)
    if [((x * y * z) < 22000 or (x * y * z) > 28000)] != [((x * y * z) < 22000 or (x * y * z) > 99000)]:
        triggered.add(132)

    # 133-134：湿度扭矩同时偏低（原135-136，编号顺延）
    if [(x < 43 and y < 160)] != [(x < 43 and y < 1680)]:
        triggered.add(133)
    if [(x < 43 and y < 160)] != [(x < 43 and y < 1600)]:
        triggered.add(134)

    # 135-136：湿度扭矩同时偏高（原137-138，编号顺延）
    if [(x > 57 and y > 190)] != [(x > 57 and y * 9 > 190)]:
        triggered.add(135)
    if [(x > 57 and y > 190)] != [(x > 57 and y * 8 > 190)]:
        triggered.add(136)

    # 137-138：湿度电流同时偏低（原139-140，编号顺延）
    if [(x < 43 and z < 11)] != [(x < 43 and z * 2 < 11)]:
        triggered.add(137)
    if [(x < 43 and z < 11)] != [(x < 43 and 8 < 11)]:
        triggered.add(138)

    # 139-140：湿度电流同时偏高（原141-142，编号顺延）
    if [(x > 57 and z > 14)] != [(x > 57 and z * 8 > 14)]:
        triggered.add(139)
    if [(x > 57 and z > 14)] != [(x > 57 and 59 > 14)]:
        triggered.add(140)

    # 141-142：扭矩电流同时偏低（原143-144，编号顺延）
    if [(y < 160 and z < 11)] != [(y < 160 and z * 8 < 11)]:
        triggered.add(141)
    if [(y < 160 and z < 11)] != [(y < 160 and 9 < 11)]:
        triggered.add(142)

    # 143-144：扭矩电流同时偏高（原145-146，编号顺延）
    if [(y > 190 and z > 14)] != [(y > 190 and z * 44 > 14)]:
        triggered.add(143)
    if [(y > 190 and z > 14)] != [(y > 190 and 23 > 14)]:
        triggered.add(144)

    # 145-146：湿度在临界范围（原147-148，编号顺延）
    if [(x < 38 or x > 62)] != [(x < 38 or x * 78 > 62)]:
        triggered.add(145)
    if [(x < 38 or x > 62)] != [(x < 38 or x > 162)]:
        triggered.add(146)

    # 147-148：扭矩在临界范围（原149-150，编号顺延）
    if [(y < 145 or y > 205)] != [(y < 145 or y > 150)]:
        triggered.add(147)
    if [(y < 145 or y > 205)] != [(y < 115 or y > 205)]:
        triggered.add(148)

    # 149-150：电流在临界范围（原151-152，编号顺延）
    if [(z < 9.5 or z > 15.5)] != [(z < 9.5 or z > 155)]:
        triggered.add(149)
    if [(z < 9.5 or z > 15.5)] != [(z < 9.5 or z > 95.5)]:
        triggered.add(150)

    # 151-152：所有参数严重偏低（原153-154，编号顺延）
    if [(x < 35 and y < 155 and z < 10.5)] != [(x < 35 and y < 155 and z < 105)]:
        triggered.add(151)
    if [(x < 35 and y < 155 and z < 10.5)] != [(x < 35 and y < 155 and z * 8 < 10.5)]:
        triggered.add(152)

    # 155-156：三维偏离过大（原157-158，编号顺延）
    if [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 16) > 100] != [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 16) > 900]:
        triggered.add(153)
    if [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 16) > 100] != [((x - 50) ** 2 + (y / 4 - 43.75) ** 2 + (z - 12.5) ** 2 * 89) > 100]:
        triggered.add(154)

    # 157-158：三元积偏离过大（原159-160，编号顺延）
    if [(abs((x * y * z) / 25000 - 1) > 0.2)] != [(abs((x * y * z) / 25000 - 1) > 2)]:
        triggered.add(155)
    if [(abs((x * y * z) / 25000 - 1) > 0.2)] != [(abs((x * y * z) / 25000 - 1) > 7)]:
        triggered.add(156)

    return triggered



target_paths = [
    {1, 2, 3, 4, 5, 6, 11, 12, 17, 21, 22, 23, 24, 29, 31, 32, 33, 34, 36, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48, 53,
     54, 55, 56, 67, 68, 69, 70, 71, 72, 77, 78, 79, 80, 95, 96, 97, 99, 100, 101, 102, 103, 105, 106, 113, 114, 117,
     118, 123, 124, 125, 126, 129, 130, 135, 136, 139, 140, 146, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 17, 22, 23, 29, 31, 32, 33, 34, 36, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48, 53, 54, 55, 56, 57,
     58, 59, 60, 67, 68, 69, 70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 101, 102, 105, 106, 113, 114, 117, 118, 123, 124,
     125, 126, 127, 128, 129, 130, 135, 136, 139, 140, 146, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 11, 12, 17, 19, 21, 22, 23, 24, 29, 31, 32, 33, 34, 36, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48,
     53, 54, 55, 56, 67, 68, 69, 70, 77, 78, 79, 80, 95, 96, 97, 99, 100, 101, 102, 103, 105, 106, 113, 114, 117, 118,
     123, 124, 125, 126, 129, 130, 135, 136, 139, 140, 145, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 11, 12, 15, 16, 17, 23, 24, 29, 31, 32, 33, 34, 36, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48, 53,
     54, 55, 56, 59, 60, 67, 68, 69, 70, 77, 78, 79, 80, 95, 96, 97, 99, 100, 101, 102, 105, 106, 113, 114, 117, 118,
     123, 124, 125, 126, 129, 130, 135, 136, 139, 140, 146, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 11, 12, 17, 22, 23, 24, 29, 31, 33, 34, 36, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48, 53, 54, 55,
     56, 67, 68, 69, 70, 71, 72, 77, 78, 79, 80, 95, 96, 97, 99, 100, 101, 102, 103, 105, 106, 113, 114, 117, 118, 123,
     124, 125, 126, 129, 130, 139, 140, 143, 144, 146, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 11, 12, 17, 22, 23, 24, 29, 31, 32, 33, 34, 35, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48, 53, 54,
     55, 56, 59, 60, 67, 68, 69, 70, 77, 78, 79, 80, 95, 96, 97, 101, 102, 105, 106, 113, 114, 117, 118, 120, 123, 124,
     125, 126, 129, 130, 135, 136, 139, 140, 142, 146, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 17, 22, 23, 29, 31, 33, 34, 37, 38, 39, 40, 41, 42, 43, 44, 47, 48, 51, 52, 53, 54, 55, 56, 69,
     70, 71, 72, 73, 74, 77, 78, 79, 80, 95, 96, 97, 99, 100, 101, 102, 103, 105, 106, 108, 113, 114, 117, 118, 123,
     124, 125, 126, 129, 130, 139, 140, 143, 144, 146, 147, 153, 156},

    {1, 2, 5, 6, 9, 10, 17, 19, 20, 22, 23, 24, 28, 29, 35, 37, 38, 41, 42, 49, 53, 54, 55, 56, 57, 58, 63, 64, 66, 68,
     71, 72, 77, 78, 79, 80, 95, 96, 101, 102, 105, 106, 109, 113, 114, 117, 118, 120, 121, 122, 123, 124, 125, 126,
     127, 128, 129, 130, 131, 132, 142, 145, 148, 153, 155, 156},

    {1, 2, 3, 4, 5, 6, 17, 19, 22, 23, 29, 33, 34, 35, 37, 38, 39, 40, 41, 42, 53, 54, 55, 56, 57, 58, 59, 60, 67, 68,
     69, 70, 73, 74, 77, 78, 79, 80, 85, 86, 95, 96, 97, 101, 102, 105, 106, 113, 114, 117, 118, 120, 123, 124, 127,
     128, 129, 130, 135, 136, 139, 140, 142, 145, 148, 153, 156},

    {1, 2, 3, 4, 17, 19, 21, 22, 29, 31, 32, 33, 34, 37, 38, 39, 40, 47, 48, 51, 52, 53, 54, 55, 56, 57, 58, 63, 64, 69,
     70, 73, 74, 77, 78, 79, 80, 95, 96, 97, 99, 100, 101, 102, 105, 106, 113, 114, 117, 118, 123, 124, 125, 126, 127,
     128, 129, 130, 135, 136, 145, 147, 149, 150, 153, 156},

    {1, 2, 3, 4, 5, 6, 18, 19, 22, 23, 25, 26, 29, 31, 36, 37, 38, 39, 40, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 55,
     56, 57, 58, 68, 69, 70, 71, 72, 77, 78, 79, 80, 84, 97, 101, 102, 103, 105, 106, 108, 117, 118, 119, 125, 126, 127,
     128, 129, 130, 139, 140, 143, 144, 145, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 22, 23, 25, 26, 29, 31, 35, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 53, 54, 56, 57, 58,
     61, 62, 68, 69, 70, 71, 72, 77, 78, 79, 80, 97, 101, 102, 105, 106, 108, 117, 118, 119, 123, 124, 125, 126, 127,
     128, 129, 130, 139, 140, 143, 144, 146, 147, 153, 156},

    {1, 2, 3, 4, 17, 19, 22, 29, 33, 34, 37, 38, 39, 40, 41, 42, 53, 54, 55, 56, 57, 58, 59, 60, 63, 64, 69, 70, 73, 74,
     77, 78, 79, 80, 87, 88, 95, 96, 97, 101, 102, 105, 106, 113, 114, 117, 118, 120, 123, 124, 125, 126, 127, 128, 129,
     130, 135, 136, 142, 145, 148, 149, 150, 153, 156},

    {1, 2, 3, 4, 5, 6, 18, 22, 23, 24, 29, 31, 36, 37, 38, 39, 40, 43, 44, 45, 46, 47, 48, 51, 52, 53, 54, 55, 56, 57,
     58, 67, 68, 69, 70, 71, 72, 77, 78, 79, 80, 83, 84, 97, 101, 102, 103, 105, 106, 108, 117, 118, 125, 126, 127, 128,
     129, 130, 139, 140, 143, 144, 145, 147, 153, 156},

    {1, 2, 5, 6, 17, 19, 20, 22, 23, 27, 29, 35, 37, 38, 39, 40, 41, 42, 49, 50, 53, 54, 55, 56, 57, 58, 63, 64, 66, 67,
     68, 69, 70, 71, 72, 77, 78, 79, 80, 95, 96, 101, 102, 105, 106, 113, 114, 117, 118, 120, 123, 124, 125, 126, 127,
     128, 129, 130, 131, 132, 142, 145, 148, 153, 156},

    {1, 2, 3, 4, 13, 14, 17, 22, 29, 31, 32, 33, 34, 37, 38, 39, 40, 51, 52, 53, 54, 55, 56, 57, 58, 63, 64, 69, 70, 75,
     76, 77, 78, 79, 80, 89, 90, 93, 94, 95, 96, 97, 99, 100, 101, 102, 105, 106, 112, 113, 114, 117, 118, 125, 126,
     127, 128, 129, 130, 145, 147, 149, 150, 153, 156},

    {1, 2, 3, 4, 7, 8, 13, 14, 17, 21, 22, 27, 29, 33, 34, 37, 38, 39, 40, 51, 52, 56, 57, 58, 63, 64, 69, 70, 77, 78,
     79, 80, 87, 88, 89, 90, 92, 93, 94, 95, 96, 97, 99, 100, 101, 102, 105, 106, 111, 112, 113, 114, 117, 118, 125,
     126, 127, 128, 129, 130, 145, 147, 149, 150, 156},

    {1, 2, 5, 6, 9, 10, 17, 19, 20, 23, 24, 28, 30, 35, 41, 42, 49, 53, 54, 55, 56, 57, 58, 63, 64, 66, 68, 71, 72, 77,
     78, 79, 80, 95, 96, 101, 102, 105, 106, 109, 110, 113, 114, 117, 118, 121, 122, 123, 124, 125, 126, 127, 128, 129,
     130, 131, 132, 142, 145, 148, 153, 155, 156},

    {1, 2, 3, 4, 5, 6, 17, 22, 23, 24, 28, 29, 35, 37, 38, 50, 53, 54, 56, 57, 58, 63, 64, 66, 68, 71, 72, 77, 78, 79,
     80, 89, 90, 95, 96, 101, 102, 105, 106, 109, 113, 114, 117, 118, 121, 122, 123, 124, 125, 126, 127, 128, 129, 130,
     131, 132, 138, 142, 148, 151, 153, 155, 156},

    {1, 2, 3, 4, 5, 6, 11, 12, 15, 16, 17, 19, 23, 24, 28, 29, 36, 37, 38, 39, 40, 49, 53, 54, 55, 56, 63, 64, 69, 70,
     79, 80, 82, 89, 90, 95, 96, 101, 102, 103, 105, 106, 113, 114, 116, 117, 118, 121, 122, 123, 124, 125, 126, 129,
     130, 131, 132, 133, 134, 138, 145, 147, 156},

    {1, 2, 3, 4, 5, 6, 17, 19, 20, 22, 23, 27, 29, 37, 38, 39, 40, 49, 50, 51, 52, 53, 54, 55, 56, 63, 64, 67, 68, 69,
     70, 77, 78, 79, 80, 87, 88, 89, 90, 95, 96, 97, 101, 102, 103, 105, 106, 113, 114, 116, 117, 118, 123, 124, 125,
     126, 129, 130, 143, 144, 145, 147, 154, 156},

    {1, 2, 3, 4, 7, 8, 19, 20, 21, 22, 25, 26, 28, 30, 35, 39, 40, 43, 44, 46, 53, 54, 55, 56, 57, 58, 61, 62, 65, 66,
     73, 74, 77, 78, 79, 80, 101, 102, 104, 105, 106, 109, 110, 117, 118, 119, 121, 122, 123, 124, 125, 126, 127, 129,
     130, 131, 132, 145, 147, 153, 155, 156},

    {1, 2, 3, 4, 5, 6, 19, 21, 22, 28, 29, 35, 37, 38, 39, 40, 43, 44, 45, 46, 53, 54, 57, 58, 61, 62, 66, 73, 74, 77,
     78, 79, 80, 101, 102, 103, 104, 105, 106, 117, 118, 121, 122, 123, 124, 125, 126, 127, 128, 129, 130, 131, 132,
     133, 134, 137, 145, 147, 153, 155, 156},

    {1, 2, 5, 6, 17, 23, 29, 33, 34, 35, 37, 38, 41, 42, 43, 44, 53, 54, 55, 56, 57, 58, 59, 60, 67, 68, 69, 70, 73, 74,
     79, 80, 95, 96, 98, 101, 102, 105, 106, 109, 113, 114, 117, 118, 120, 123, 124, 127, 128, 129, 130, 131, 132, 135,
     136, 139, 140, 142, 146, 153, 156},

    {1, 2, 3, 4, 5, 6, 22, 25, 26, 29, 31, 32, 35, 37, 38, 41, 42, 43, 44, 45, 46, 47, 48, 50, 53, 54, 56, 57, 58, 59,
     60, 69, 70, 79, 80, 101, 102, 105, 106, 109, 117, 118, 123, 124, 125, 126, 127, 128, 129, 130, 131, 132, 135, 136,
     139, 140, 141, 146, 147, 153, 156},

    {1, 2, 17, 19, 27, 29, 35, 37, 38, 39, 40, 41, 42, 49, 50, 53, 54, 56, 57, 58, 63, 64, 66, 67, 68, 72, 77, 78, 79,
     80, 95, 96, 101, 102, 105, 106, 107, 109, 113, 114, 117, 118, 123, 124, 125, 126, 127, 129, 130, 131, 132, 138,
     142, 145, 149, 150, 153, 155, 156},

    {1, 2, 3, 4, 5, 6, 7, 8, 13, 14, 15, 16, 23, 25, 26, 28, 29, 35, 37, 38, 39, 40, 43, 44, 45, 46, 55, 56, 57, 58, 66,
     69, 70, 77, 78, 79, 80, 81, 92, 103, 104, 105, 106, 117, 118, 119, 121, 122, 125, 126, 127, 128, 129, 130, 131,
     132, 145, 147, 153, 156},

    {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 17, 19, 20, 21, 22, 23, 24, 28, 29, 36, 37, 38, 39, 40, 49, 50, 53, 54, 67,
     68, 69, 70, 87, 88, 95, 96, 98, 101, 102, 103, 105, 106, 113, 114, 115, 116, 117, 118, 121, 122, 129, 130, 145,
     147, 156},

    {1, 2, 3, 4, 5, 6, 7, 8, 13, 14, 17, 23, 27, 29, 36, 37, 38, 39, 40, 49, 50, 63, 64, 67, 68, 69, 70, 79, 80, 84, 85,
     86, 87, 88, 89, 90, 91, 92, 95, 96, 97, 101, 102, 103, 105, 106, 113, 114, 116, 117, 118, 129, 130, 145, 147, 154,
     156},

    {3, 4, 5, 6, 22, 28, 30, 35, 45, 46, 53, 54, 55, 56, 63, 64, 65, 66, 71, 72, 73, 74, 77, 78, 79, 80, 103, 109, 110,
     117, 118, 121, 122, 123, 124, 125, 126, 127, 128, 131, 132, 137, 141, 147, 152, 153, 155, 156},
]


def jaccard_similarity(set1: Set[int], set2: Set[int]) -> float:
    """If set1 covers target set2, the match score is treated as 1."""
    if set2.issubset(set1):
        return 1.0
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    return intersection / union if union != 0 else 0.0


def compute_reward(state, target_path, triggered):
    """Compute the DQN reward."""
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward


class SimpleReplayBuffer:
    def __init__(self, capacity=5000):
        self.buffer = deque(maxlen=capacity)

    def append(self, experience):
        self.buffer.append(experience)

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return random.sample(self.buffer, len(self.buffer))
        return random.sample(self.buffer, batch_size)

    def __len__(self):
        return len(self.buffer)


class SimpleDQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super().__init__()
        self.fc1 = nn.Linear(state_dim, 64)
        self.fc2 = nn.Linear(64, 32)
        self.fc3 = nn.Linear(32, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)


class SimpleDQNAgent:
    def __init__(self, state_dim, action_dim, learning_rate=0.001):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.epsilon = 0.3
        self.gamma = 0.99

        self.model = SimpleDQN(state_dim, action_dim).to(device)
        self.target_model = SimpleDQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())
        self.replay_buffer = SimpleReplayBuffer(capacity=5000)

    def decode_action(self, action_idx):
        """Decode the action."""
        light_deltas = [int((LIGHT_MAX - LIGHT_MIN) * p) for p in [0.05, 0.1, 0.2, 0.5, 0.7, -0.05, -0.1, -0.2, -0.5, -0.7]]
        moisture_deltas = [int((MOISTURE_MAX - MOISTURE_MIN) * p) for p in [0.05, 0.1, 0.2, 0.5, 0.7, -0.05, -0.1, -0.2, -0.5, -0.7]]
        temp_deltas = [int((TEMP_MAX - TEMP_MIN) * p) for p in [0.05, 0.1, 0.2, 0.5, 0.7, -0.05, -0.1, -0.2, -0.5, -0.7]]

        dim = action_idx // 10
        delta_idx = action_idx % 10

        if dim == 0:
            return (light_deltas[delta_idx], 0, 0)
        if dim == 1:
            return (0, moisture_deltas[delta_idx], 0)
        return (0, 0, temp_deltas[delta_idx])

    def store_transition(self, state, action, reward, next_state, done):
        self.replay_buffer.append((state, action, reward, next_state, done))

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return

        batch = self.replay_buffer.sample(batch_size)
        states, actions, rewards, next_states, dones = zip(*batch)

        states = torch.tensor(np.array(states), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array(next_states), dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))

        loss = nn.MSELoss()(current_q_values, target_q_values)

        self.optimizer.zero_grad()
        loss.backward()
        self.optimizer.step()

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())

    def get_best_samples(self, target_path, num_samples=20):
        """Screen DQN candidate samples from the replay buffer. This stage is counted separately in T_sample_screening."""
        if len(self.replay_buffer) == 0:
            return []

        samples_with_scores = []
        for state_norm, _, _, _, _ in self.replay_buffer.buffer:
            state_original = denormalize_state(state_norm)
            triggered = execute_Tr(state_original)
            reward = compute_reward(state_original, target_path, triggered)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_scores.append((state_original, reward, sim, triggered))

        samples_with_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_scores[:num_samples]


def train_dqn_for_path(path_idx, target_path, num_samples=200):
    """Train a standard DQN for a single path."""
    print(f"  Start training path {path_idx + 1} DQN model...")
    agent = SimpleDQNAgent(state_dim=3, action_dim=30)

    random_states = [normalize_state(generate_random_state()) for _ in range(num_samples)]

    STEPS_PER_SAMPLE = 3
    EPOCHS = 5
    BATCH_SIZE = 32

    step_count = 0
    for _ in range(EPOCHS):
        for state_norm in random_states:
            state_norm = tuple(state_norm)
            state_original = denormalize_state(state_norm)

            for step in range(STEPS_PER_SAMPLE):
                legal_actions = []
                for a in range(agent.action_dim):
                    dx, dy, dz = agent.decode_action(a)
                    next_state_candidate = (state_original[0] + dx, state_original[1] + dy, state_original[2] + dz)
                    if is_state_valid(next_state_candidate):
                        legal_actions.append(a)

                if not legal_actions:
                    break

                if random.random() < agent.epsilon:
                    action = random.choice(legal_actions)
                else:
                    state_tensor = torch.tensor(state_norm, dtype=torch.float32).unsqueeze(0).to(device)
                    with torch.no_grad():
                        q_values = agent.model(state_tensor)[0]
                    action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                dx, dy, dz = agent.decode_action(action)
                next_state_original = (state_original[0] + dx, state_original[1] + dy, state_original[2] + dz)
                next_state_norm = normalize_state(next_state_original)

                triggered = execute_Tr(next_state_original)
                reward = compute_reward(next_state_original, target_path, triggered)
                done = (step == STEPS_PER_SAMPLE - 1)

                agent.store_transition(state_norm, action, reward, next_state_norm, done)

                state_norm = next_state_norm
                state_original = next_state_original
                step_count += 1

                if step_count % 50 == 0 and len(agent.replay_buffer) >= BATCH_SIZE:
                    agent.train(BATCH_SIZE)

                if step_count % 100 == 0:
                    agent.update_target_model()

    print(f"  Path {path_idx + 1} training completed, replay buffer size: {len(agent.replay_buffer)}")
    return agent


class Particle:
    def __init__(self, initial_position=None):
        if initial_position is not None:
            self.position = clip_state(np.array(initial_position, dtype=float))
        else:
            self.position = np.array([
                np.random.uniform(BOUNDS["light"][0], BOUNDS["light"][1]),
                np.random.uniform(BOUNDS["moisture"][0], BOUNDS["moisture"][1]),
                np.random.uniform(BOUNDS["temp"][0], BOUNDS["temp"][1]),
            ], dtype=float)

        spans = np.array([
            BOUNDS["light"][1] - BOUNDS["light"][0],
            BOUNDS["moisture"][1] - BOUNDS["moisture"][0],
            BOUNDS["temp"][1] - BOUNDS["temp"][0],
        ], dtype=float)
        self.velocity = np.random.uniform(-0.1 * spans, 0.1 * spans)
        self.best_position = self.position.copy()
        self.best_fitness = -1.0
        self.fitness = -1.0


class BasicPSO:
    """
    Standard PSO: contains only standard velocity and position updates.
    Mutation coefficients, opposite particles, local-best reset, and other improved PSO mechanisms have been removed.
    """

    def __init__(self, target_path, swarm_size=20, dqn_samples=None):
        self.target_path = target_path
        self.swarm_size = swarm_size
        self.particles = []
        self.global_best_position = None
        self.global_best_fitness = -1.0

        if dqn_samples is not None and len(dqn_samples) > 0:
            num_direct = min(len(dqn_samples), swarm_size)
            for i in range(num_direct):
                state_tuple, _, _, _ = dqn_samples[i]
                self.particles.append(Particle(initial_position=state_tuple))

            while len(self.particles) < swarm_size:
                base_idx = len(self.particles) % len(dqn_samples)
                state_tuple, _, _, _ = dqn_samples[base_idx]
                perturbed = np.array(state_tuple, dtype=float) + np.random.randint(-10, 11, size=3)
                self.particles.append(Particle(initial_position=clip_state(perturbed)))
        else:
            self.particles = [Particle() for _ in range(swarm_size)]

        for particle in self.particles:
            particle.fitness = self.fitness_function(particle.position)
            particle.best_fitness = particle.fitness
            particle.best_position = particle.position.copy()

            if particle.fitness > self.global_best_fitness:
                self.global_best_fitness = particle.fitness
                self.global_best_position = particle.position.copy()

    def fitness_function(self, position):
        try:
            triggered = execute_Tr(position)
            return jaccard_similarity(triggered, self.target_path)
        except Exception:
            return 0.0

    def update(self):
        spans = np.array([
            BOUNDS["light"][1] - BOUNDS["light"][0],
            BOUNDS["moisture"][1] - BOUNDS["moisture"][0],
            BOUNDS["temp"][1] - BOUNDS["temp"][0],
        ], dtype=float)
        max_velocity = PSO_VMAX_RATIO * spans

        for particle in self.particles:
            r1 = np.random.random(3)
            r2 = np.random.random(3)

            particle.velocity = (
                PSO_W * particle.velocity
                + PSO_C1 * r1 * (particle.best_position - particle.position)
                + PSO_C2 * r2 * (self.global_best_position - particle.position)
            )
            particle.velocity = np.clip(particle.velocity, -max_velocity, max_velocity)

            particle.position = clip_state(particle.position + particle.velocity)
            particle.fitness = self.fitness_function(particle.position)

            if particle.fitness > particle.best_fitness:
                particle.best_fitness = particle.fitness
                particle.best_position = particle.position.copy()

            if particle.fitness > self.global_best_fitness:
                self.global_best_fitness = particle.fitness
                self.global_best_position = particle.position.copy()


def _set_header(ws, headers, widths, header_fill, header_font, border, center_align):
    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_row(ws, row_idx, row_data, border, alignment, fill=None):
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.border = border
        cell.alignment = alignment
        if fill is not None:
            cell.fill = fill


def export_time_metrics_to_excel(all_run_results, experiment_total_time=None, filename=None):
    """Export only timing metrics and per-path iteration counts."""
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"DQN_BasicPSO_Time_Metrics_{timestamp}.xlsx"

    output_dir = os.path.join(os.getcwd(), "results")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    num_runs = len(all_run_results)
    num_paths = len(target_paths)

    # Worksheet 1: runtime summary
    ws1 = wb.active
    ws1.title = "Runtime Summary"
    ws1.sheet_view.showGridLines = False
    headers1 = [
        "Run", "Number of Paths", "DQNTraining Total Time(s)", "Sample Screening Total Time(s)", "DQNDirect Check Total Time(s)",
        "PSOInitialization Total Time(s)", "PSOSearch Total Time(s)", "Total Algorithm Time(s)", "Total Iterations", "Average Iterations per Path"
    ]
    widths1 = [12, 10, 18, 18, 22, 18, 18, 16, 14, 18]
    _set_header(ws1, headers1, widths1, header_fill, header_font, border, center_align)

    for run_idx, run_result in enumerate(all_run_results, 1):
        path_results = run_result["paths"]
        total_dqn_train = sum(p["T_DQN_train"] for p in path_results)
        total_screen = sum(p["T_sample_screening"] for p in path_results)
        total_direct_check = sum(p["T_direct_check"] for p in path_results)
        total_pso_init = sum(p["T_PSO_init"] for p in path_results)
        total_pso_search = sum(p["T_PSO_search"] for p in path_results)
        total_iterations = sum(p["iterations"] for p in path_results)

        row_data = [
            f" {run_idx}",
            num_paths,
            round(total_dqn_train, 6),
            round(total_screen, 6),
            round(total_direct_check, 6),
            round(total_pso_init, 6),
            round(total_pso_search, 6),
            round(run_result["T_run_algorithm_total"], 6),
            int(total_iterations),
            round(total_iterations / num_paths, 2),
        ]
        _write_row(ws1, run_idx + 1, row_data, border, center_align, alternate_fill if run_idx % 2 == 0 else None)

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:J{num_runs + 1}"

    # Worksheet 2: path timing and iteration details
    ws2 = wb.create_sheet(title="Path Timing and Iteration Details")
    ws2.sheet_view.showGridLines = False
    headers2 = [
        "Run", "Path ID", "DQN training(s)", "Sample Screening Time(s)", "DQNDirect Check Time(s)",
        "PSOInitialization Time(s)", "PSOSearch Time(s)", "Path Total Algorithm Time(s)", "Iterations"
    ]
    widths2 = [12, 12, 18, 18, 22, 18, 18, 18, 12]
    _set_header(ws2, headers2, widths2, header_fill, header_font, border, center_align)

    row_idx = 2
    for run_idx, run_result in enumerate(all_run_results, 1):
        for path_result in run_result["paths"]:
            row_data = [
                f" {run_idx}",
                f"Path  {path_result['path_idx'] + 1}",
                round(path_result["T_DQN_train"], 6),
                round(path_result["T_sample_screening"], 6),
                round(path_result["T_direct_check"], 6),
                round(path_result["T_PSO_init"], 6),
                round(path_result["T_PSO_search"], 6),
                round(path_result["T_path_algorithm_total"], 6),
                int(path_result["iterations"]),
            ]
            _write_row(ws2, row_idx, row_data, border, center_align, alternate_fill if row_idx % 2 == 0 else None)
            row_idx += 1

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:I{row_idx - 1}"

    # Worksheet 3: path aggregate statistics
    ws3 = wb.create_sheet(title="Path Aggregate Statistics")
    ws3.sheet_view.showGridLines = False
    headers3 = [
        "Path ID", "Average DQN Training Time(s)", "Average Sample Screening Time(s)", "Average PSO Initialization Time(s)",
        "Average PSO Search Time(s)", "Average Path Total Algorithm Time(s)", "Average Iterations", "Minimum Iterations", "Maximum Iterations"
    ]
    widths3 = [12, 22, 22, 22, 22, 22, 16, 14, 14]
    _set_header(ws3, headers3, widths3, header_fill, header_font, border, center_align)

    for path_idx in range(num_paths):
        records = [run_result["paths"][path_idx] for run_result in all_run_results]
        iterations = [r["iterations"] for r in records]
        row_data = [
            f"Path  {path_idx + 1}",
            round(float(np.mean([r["T_DQN_train"] for r in records])), 6),
            round(float(np.mean([r["T_sample_screening"] for r in records])), 6),
            round(float(np.mean([r["T_PSO_init"] for r in records])), 6),
            round(float(np.mean([r["T_PSO_search"] for r in records])), 6),
            round(float(np.mean([r["T_path_algorithm_total"] for r in records])), 6),
            round(float(np.mean(iterations)), 2),
            int(np.min(iterations)),
            int(np.max(iterations)),
        ]
        _write_row(ws3, path_idx + 2, row_data, border, center_align, alternate_fill if (path_idx + 1) % 2 == 0 else None)

    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:I{num_paths + 1}"

    # Worksheet 4: experiment overview
    ws4 = wb.create_sheet(title="Experiment Overview")
    ws4.sheet_view.showGridLines = False
    headers4 = ["Metric", "Value"]
    widths4 = [32, 24]
    _set_header(ws4, headers4, widths4, header_fill, header_font, border, center_align)

    all_path_records = [p for run_result in all_run_results for p in run_result["paths"]]
    all_iterations = [p["iterations"] for p in all_path_records]
    overview_rows = [
        ["Experiment Runs", num_runs],
        ["Number of Target Paths", num_paths],
        ["Total Experiment Wall-clock Time(s)", round(experiment_total_time, 6) if experiment_total_time is not None else ""],
        ["Average Total Algorithm Time per Run(s)", round(float(np.mean([r["T_run_algorithm_total"] for r in all_run_results])), 6)],
        ["Average Total Algorithm Time per Path(s)", round(float(np.mean([p["T_path_algorithm_total"] for p in all_path_records])), 6)],
        ["Average DQN Training Time per Path(s)", round(float(np.mean([p["T_DQN_train"] for p in all_path_records])), 6)],
        ["Average Sample Screening Time per Path(s)", round(float(np.mean([p["T_sample_screening"] for p in all_path_records])), 6)],
        ["Average PSO Search Time per Path(s)", round(float(np.mean([p["T_PSO_search"] for p in all_path_records])), 6)],
        ["Average Iterations per Path", round(float(np.mean(all_iterations)), 2)],
        ["Maximum Iterations per Path", int(np.max(all_iterations))],
        ["Minimum Iterations per Path", int(np.min(all_iterations))],
    ]

    for idx, row_data in enumerate(overview_rows, 2):
        _write_row(ws4, idx, row_data, border, center_align, alternate_fill if idx % 2 == 0 else None)

    wb.save(filepath)
    print(f"\n Timing metrics exported to: {filepath}")
    return filepath


def run_single_experiment(run_num, max_iterations=3000):
    """Run one DQN + standard PSO ."""
    print(f"\n{'=' * 80}")
    print(f"Start run  {run_num}  run: DQN + standard PSO ")
    print(f"{'=' * 80}")

    run_start = time.perf_counter()
    path_results = []

    for path_idx, target_path in enumerate(target_paths):
        print(f"\nRun {run_num} run - Path {path_idx + 1}")
        path_total_start = time.perf_counter()

        # 1) DQN training
        dqn_train_start = time.perf_counter()
        agent = train_dqn_for_path(path_idx, target_path, num_samples=200)
        T_DQN_train = time.perf_counter() - dqn_train_start

        # 2) DQNSample Screening Time
        screening_start = time.perf_counter()
        dqn_samples = agent.get_best_samples(target_path, num_samples=20)
        T_sample_screening = time.perf_counter() - screening_start

        # 3) DQN
        direct_check_start = time.perf_counter()
        direct_solution_found = False
        for state_tuple, reward, sim, triggered in dqn_samples:
            if sim >= 1.0 or target_path.issubset(triggered):
                direct_solution_found = True
                break
        T_direct_check = time.perf_counter() - direct_check_start

        T_PSO_init = 0.0
        T_PSO_search = 0.0
        iterations_used = 0

        if direct_solution_found:
            print(
                f"  Path {path_idx + 1}: DQN | "
                f"DQN training {T_DQN_train:.6f}s | screening {T_sample_screening:.6f}s | iterations {iterations_used}"
            )
        else:
            # 4) standard PSOInitialization Time
            pso_init_start = time.perf_counter()
            pso = BasicPSO(target_path, swarm_size=20, dqn_samples=dqn_samples)
            T_PSO_init = time.perf_counter() - pso_init_start

            # 5) standard PSOSearch Time
            pso_search_start = time.perf_counter()
            iterations_used = max_iterations
            for iteration in range(max_iterations):
                pso.update()
                if pso.global_best_fitness >= 1.0:
                    iterations_used = iteration + 1
                    break
            T_PSO_search = time.perf_counter() - pso_search_start

            print(
                f"  Path {path_idx + 1}: "
                f"DQN training {T_DQN_train:.6f}s | screening {T_sample_screening:.6f}s | "
                f"PSO {T_PSO_init:.6f}s | PSO {T_PSO_search:.6f}s | "
                f"iterations {iterations_used}"
            )

        T_path_algorithm_total = time.perf_counter() - path_total_start

        path_results.append({
            "path_idx": path_idx,
            "T_DQN_train": T_DQN_train,
            "T_sample_screening": T_sample_screening,
            "T_direct_check": T_direct_check,
            "T_PSO_init": T_PSO_init,
            "T_PSO_search": T_PSO_search,
            "T_path_algorithm_total": T_path_algorithm_total,
            "iterations": iterations_used,
        })

    T_run_algorithm_total = time.perf_counter() - run_start
    print(f"\nRun {run_num} runcompleted | Total Algorithm Time {T_run_algorithm_total:.6f}s")
    return {
        "run_num": run_num,
        "paths": path_results,
        "T_run_algorithm_total": T_run_algorithm_total,
    }


def run_multiple_experiments(num_runs):
    print("\n" + "=" * 80)
    print(f"DQN + standard PSO - {num_runs}: Metric")
    print("=" * 80)
    print(f"Number of Target Paths: {len(target_paths)}")
    print("Statistics: DQN training, Sample Screening Time, PSOInitialization Time, PSOSearch Time, Path Iterations")
    print("=" * 80)

    all_run_results = []
    experiment_start = time.perf_counter()

    for run_num in range(1, num_runs + 1):
        run_result = run_single_experiment(run_num)
        all_run_results.append(run_result)

    experiment_total_time = time.perf_counter() - experiment_start
    print(f"\n{'=' * 80}")
    print(f"All {num_runs} runcompleted | Total Experiment Wall-clock Time {experiment_total_time:.6f}s")
    print(f"{'=' * 80}\n")
    return all_run_results, experiment_total_time


if __name__ == "__main__":
    print("=" * 80)
    print("DQN + standard PSO - MetricPath Iterations")
    print("=" * 80)
    print(f"Current configuration: Run = {NUM_RUNS}")
    print(f"Number of Paths: {len(target_paths)}")
    print(f"Device: {device}")
    print("=" * 80)

    if len(sys.argv) > 1:
        try:
            NUM_RUNS = int(sys.argv[1])
            print(f"Read from command line: Run = {NUM_RUNS}")
        except ValueError:
            print(f"Invalid command-line argument, using default number of runs {NUM_RUNS}")

    all_results, total_time = run_multiple_experiments(num_runs=NUM_RUNS)
    export_time_metrics_to_excel(all_results, total_time)
    print("\nProgram completed")
