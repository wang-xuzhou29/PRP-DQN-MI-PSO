import os
import random
import numpy as np
import time
import psutil
import math
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
import math
NUM_RUNS = 20  # 默认实验运行次数

# === 新的三维范围设置 ===
LIGHT_MIN = 1
LIGHT_MAX = 300
MOISTURE_MIN = 1
MOISTURE_MAX = 200
TEMP_MIN = 1
TEMP_MAX = 5

BOUNDS = {
    'light': (LIGHT_MIN, LIGHT_MAX),
    'temp': (TEMP_MIN, TEMP_MAX),
    'moisture': (MOISTURE_MIN, MOISTURE_MAX)
}


def clip_state(state):
    """将状态限制在各维度的边界内"""
    return np.array([
        np.clip(state[0], BOUNDS['light'][0], BOUNDS['light'][1]),
        np.clip(state[1], BOUNDS['temp'][0], BOUNDS['temp'][1]),
        np.clip(state[2], BOUNDS['moisture'][0], BOUNDS['moisture'][1])
    ])


def execute_Tr(position):
    """执行目标函数并返回触发的路径"""
    x = int(np.clip(position[0], BOUNDS['light'][0], BOUNDS['light'][1]))
    temp = int(np.clip(position[1], BOUNDS['temp'][0], BOUNDS['temp'][1]))
    z = int(np.clip(position[2], BOUNDS['moisture'][0], BOUNDS['moisture'][1]))
    return category1_multivariable_control(x, temp, z)


# === 目标函数 ===
def category1_multivariable_control(x, y, z):
    b = {}  # 使用字典存储触发的规则

    if (30 < x < 170) != (30 < x < 100): b[0] = 1
    if (30 < x < 170) != (30 < x < 190): b[1] = 2
    if (30 < x < 170) != (30 < x < 140): b[2] = 3
    if (30 < x < 170) != (30 < x < 150): b[3] = 4
    if (30 < x < 170) != (30 < x < 160): b[4] = 5
    if (100 < y < 200) != (150 < y < 200): b[5] = 6
    if (100 < y < 200) != (50 < y < 200): b[6] = 7
    if (100 < y < 200) != (100 < y < 150): b[7] = 8
    if (1 < z < 4) != (1 < z < 3): b[8] = 9
    if (x > 10 and x < 152) != (x > 10 and x < 100): b[9] = 10

    # 规则11-20
    if (x > 40 and x < 152) != (x > 40 and x < 202): b[10] = 11
    if (abs(x - 50) < 4) != (abs(x - 50) < 2): b[11] = 12
    if (abs(x - 50) < 4) != (abs(x + 50) < 4): b[12] = 13
    if (abs(x - 50) < 4) != (abs(x * 50) < 4): b[13] = 14
    if (x > 135) != (x > 215): b[14] = 15
    if (x > 135) != (x > 235): b[15] = 16
    if (x > 135) != (x > 225): b[16] = 17
    if (y > 100) != (y > 120): b[17] = 18
    if (y > 100) != (y > 140): b[18] = 19
    if (y > 100) != (y > 50): b[19] = 20

    # 规则21-30
    if (x < 165) != (x < 265): b[20] = 21
    if (x < 165) != (x < 245): b[21] = 22
    if (x < 165) != (x < 262): b[22] = 23
    if (y < 150) != (y < 130): b[23] = 24
    if (y < 150) != (x < 100): b[24] = 25
    if (y < 190) != (y < 120): b[25] = 26
    if ((x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 125) != (
            (x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 115): b[26] = 27
    if ((x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 125) != (
            (x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 85): b[27] = 28
    if ((x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 125) != (
            (x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 25): b[28] = 29
    if (x / (y / 30) > 1 and x / (y / 30) < 2) != (x / (y / 30) > 1 and x / (y / 30) < 4): b[29] = 30

    # 规则31-40
    if (x / (y / 30) > 1 and x / (y / 30) < 2) != (x / (y / 30) > 1 and x / (y / 30) < 3): b[30] = 31
    if (x / (z / 2) > 2 and x / (z / 2) < 4) != (x / (z / 2) > 1 and x / (z / 2) < 4): b[31] = 32
    if (x / (z / 2) > 2 and x / (z / 2) < 4) != (x / (z / 3) > 2 and x / (z / 2) < 4): b[32] = 33
    if (x / (z / 2) > 2 and x / (z / 2) < 4) != (x / (z / 2) > 2 and x / (z / 3) < 4): b[33] = 34
    if ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 55) != ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 155): b[
        34] = 35
    if ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 55) != ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 125): b[
        35] = 36
    if ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 55) != ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 25): b[
        36] = 37
    if (x - y / 30 > -10 and x - y / 30 < 10) != (x - y / 30 > -10 and x - y / 30 < 5): b[37] = 38
    if (x - y / 30 > -10 and x - y / 30 < 10) != (x - y / 30 > -10 and x - y / 30 < 6): b[38] = 39
    if (x - y / 30 > -10 and x - y / 30 < 10) != (x - y / 30 > -10 and x - y / 30 < 8): b[39] = 40

    # 规则41-50
    if (abs(x / (y / 30) - 1.0) < 1) != (abs(x / (y / 30) + 1.0) < 1): b[40] = 41
    if (abs(x / (y / 30) - 1.0) < 1) != (abs(x / (y / 30) - 1.0) < 2): b[41] = 42
    if (abs(x / (y / 30) - 1.0) < 1) != (abs(x / (y / 30) - 1.0) < 3): b[42] = 43
    if (x / ((y / 30) + 20) > 1 and x / ((y / 30) + 20) < 2) != (x / ((y / 30) + 20) > 1 and x / ((y / 20) + 20) < 2):
        b[43] = 44
    if (x / ((y / 30) + 20) > 1 and x / ((y / 30) + 20) < 2) != (x / ((y / 30) + 20) > 1 and x / ((y / 40) + 20) < 2):
        b[44] = 45
    if (x / ((y / 30) + 20) > 1 and x / ((y / 30) + 20) < 2) != (x / ((y / 30) + 20) > 1 and x / ((y / 30) + 40) < 2):
        b[45] = 46
    if ((z / 50) / (x / 50) > 1 and (z / 50) / (x / 50) < 3) != ((z / 50) / (x / 50) > 1 and (z / 50) / (x / 50) < 4):
        b[46] = 47
    if ((z / 50) / (x / 50) > 1 and (z / 50) / (x / 50) < 3) != ((z / 50) / (x / 50) > 2 and (z / 50) / (x / 50) < 3):
        b[47] = 48
    if ((z / 50) / (x / 50) > 1 and (z / 50) / (x / 50) < 3) != ((z / 50) / (x / 50) > 3 and (z / 50) / (x / 50) < 3):
        b[48] = 49
    if ((x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 45) != (
            (x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 70): b[49] = 50

    # 规则51-60
    if ((x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 45) != (
            (x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 60): b[50] = 51
    if ((x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 45) != (
            (x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 6 < 45): b[51] = 52
    if (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 8) != (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 48): b[
        52] = 53
    if (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 8) != (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 50): b[
        53] = 54
    if (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 8) != (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 58): b[
        54] = 55
    if ((x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 200) != (
            (x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 400): b[55] = 56
    if ((x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 200) != (
            (x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 300): b[56] = 57
    if ((x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 200) != (
            (x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 500): b[57] = 58
    if (abs((x + y / 30) - 100) < 15) != (abs((x + y / 30) - 100) < 35): b[58] = 59
    if (abs((x + y / 30) - 100) < 15) != (abs((x + y / 30) - 100) < 45): b[59] = 60

    # 规则61-70
    if (abs((x + y / 30) - 100) < 15) != (abs((x + y / 30) - 100) < 10): b[60] = 61
    if (x / 50 > 1 and x / 50 < 4) != (x / 50 > 3 and x / 50 < 4): b[61] = 62
    if (x / 50 > 1 and x / 50 < 4) != (x / 50 > 1 and x / 50 < 5): b[62] = 63
    if (x < 135) != (x < 215): b[63] = 64
    if (x < 135) != (x < 225): b[64] = 65
    if (x < 135) != (x < 225): b[65] = 66
    if (x > 65) != (x > 165): b[66] = 67
    if (x > 65) != (x > 215): b[67] = 68
    if (x > 65) != (x > 251): b[68] = 69
    if (y < 110) != (y < 80): b[69] = 70

    # 规则71-80
    if (y < 110) != (y < 50): b[70] = 71
    if (y < 110) != (y < 119): b[71] = 72
    if (x < 32 or x > 68) != (x < 32 or x > 168): b[72] = 73
    if (x < 32 or x > 68) != (x < 32 or x > 118): b[73] = 74
    if (x < 32 or x > 68) != (x < 32 or x > 218): b[74] = 75
    if (x < 35 and y < 120) != (x < 135 and y < 120): b[75] = 76
    if (x < 35 and y < 120) != (x < 235 and y < 120): b[76] = 77
    if (x < 35 and y < 120) != (x < 35 and y < 60): b[77] = 78
    if (abs(x / (y / 30) - 1.0) > 1) != (abs(x / (y / 30) - 3.0) > 1): b[78] = 79
    if (abs(x / (y / 30) - 1.0) > 1) != (abs(x / (y / 30) - 1.0) > 2): b[79] = 80

    # 规则81-90
    if (abs(x / (y / 30) - 1.0) > 1) != (abs(x / (y / 30) - 1.0) > 3): b[80] = 81
    if ((x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 135) != (
            (x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 235): b[81] = 82
    if ((x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 135) != (
            (x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 335): b[82] = 83
    if ((x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 135) != (
            (x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 435): b[83] = 84
    if (x < 30 or x > 70) != (x < 30 or x * 7 > 70): b[84] = 85
    if (x < 30 or x > 70) != (x < 30 or x > 270): b[85] = 86
    if (x < 30 or x > 70) != (x < 30 or x > 170): b[86] = 87
    if (x / (y / 30) > 1 and x / (y / 30) < 3) != (x / (y / 30) > 1 and x / (y / 20) < 3): b[87] = 88
    if (x / (y / 30) > 1 and x / (y / 30) < 3) != (x / (y / 30) > 1 and x / (y / 30) < 4): b[88] = 89
    if (x / (y / 30) > 1 and x / (y / 30) < 3) != (x / (y / 20) > 1 and x / (y / 30) < 3): b[89] = 90

    # 规则91-100
    if ((x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 112) != (
            (x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 62): b[90] = 91
    if ((x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 112) != (
            (x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 92): b[91] = 92
    if ((x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 112) != (
            (x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 82): b[92] = 93
    if (abs((x + y / 30) / 2 - 50) < 1) != (abs((x + y / 30) / 2 - 50) < 4): b[93] = 94
    if (abs((x + y / 30) / 2 - 50) < 1) != (abs((x + y / 30) / 2 + 50) < 1): b[94] = 95
    if (abs((x + y / 30) / 2 - 50) < 1) != (abs((x + y / 20) / 2 + 50) < 1): b[95] = 96
    if (x / (z / 2) > 2 and x / (z / 2) < 4) != (x / (z / 2) > 1 and x / (z / 2) < 4): b[96] = 97
    if (x / (z / 2) > 20 and x / (z / 2) < 100) != (x / (z / 3) > 20 and x / (z / 2) < 100): b[97] = 98
    if (x / (z / 2) > 20 and x / (z / 2) < 100) != (x / (z / 2) > 20 and x / (z / 2) < 150): b[98] = 99
    if ((y / 30) / (z / 50) > 38 and (y / 30) / (z / 50) < 42) != (
            (y / 30) / (z / 50) > 48 and (y / 30) / (z / 50) < 42): b[99] = 100

    # 规则101-110
    if ((y / 30) / (z / 50) > 38 and (y / 30) / (z / 50) < 42) != (
            (y / 30) / (z / 50) > 58 and (y / 30) / (z / 50) < 42): b[100] = 101
    if ((y / 30) / (z / 50) > 38 and (y / 30) / (z / 50) < 42) != (
            (y / 30) / (z / 50) > 38 and (y / 30) / (z / 50) < 72): b[101] = 102
    if ((x - 40) / 20 > 1 and (x - 40) / 20 < 6) != ((x - 40) / 30 > 1 and (x - 40) / 20 < 7): b[102] = 103
    if ((x - 40) / 20 > 1 and (x - 40) / 20 < 6) != ((x - 50) / 20 > 1 and (x - 40) / 20 < 7): b[103] = 104
    if ((x - 40) / 20 > 1 and (x - 40) / 20 < 6) != ((x - 40) / 20 > 1 and (x - 50) / 20 < 7): b[104] = 105

    triggered = set(b.values())
    return triggered

# 目标路径定义
targetPaths = [
    {1, 3, 4, 5, 6, 9, 11, 15, 16, 17, 18, 19, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 67, 68, 69, 73, 75, 77,
     82, 83, 84, 86, 87, 99, 102, 103, 104, 105},
    {1, 3, 4, 5, 6, 9, 11, 15, 16, 17, 18, 19, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 70, 71, 75, 77,
     82, 83, 84, 86, 87, 99, 102, 103, 104, 105},
    {1, 3, 4, 5, 6, 11, 15, 16, 17, 18, 19, 21, 22, 23, 25, 35, 36, 50, 52, 64, 65, 66, 68, 69, 70, 71, 73, 75, 77, 82,
     83, 84, 86, 87, 98, 103, 104, 105},
    {1, 3, 4, 5, 6, 11, 15, 16, 17, 19, 21, 22, 23, 24, 25, 26, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 73, 75, 82, 83,
     84, 86, 87, 102, 103, 104, 105},
    {1, 3, 4, 5, 7, 11, 15, 16, 17, 20, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 71, 73, 75, 77, 82, 83,
     84, 86, 87, 98, 103, 104, 105},
    {2, 6, 9, 11, 15, 16, 17, 18, 19, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 70, 71, 75, 77, 82, 83,
     84, 86, 99, 102, 103, 104, 105},
    {2, 7, 11, 15, 16, 17, 20, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 70, 71, 75, 77, 82, 83, 84, 86,
     87, 100, 101, 103, 104, 105},
    {1, 3, 4, 5, 6, 11, 15, 16, 17, 18, 19, 25, 35, 36, 50, 51, 52, 64, 65, 66, 67, 68, 69, 72, 73, 75, 77, 82, 83, 84,
     86, 87, 98, 103, 104, 105},
    {1, 3, 4, 5, 8, 9, 11, 15, 16, 17, 21, 22, 23, 26, 35, 36, 50, 51, 52, 64, 65, 66, 67, 68, 69, 73, 75, 82, 83, 84,
     86, 87, 99, 103, 104, 105},
    {1, 3, 7, 9, 10, 15, 16, 17, 20, 25, 35, 36, 50, 51, 52, 60, 62, 64, 65, 66, 67, 68, 69, 70, 71, 73, 75, 77, 82, 83,
     84, 86, 87, 102},
    {1, 6, 10, 18, 19, 25, 35, 36, 50, 51, 52, 59, 60, 62, 67, 68, 69, 70, 71, 73, 75, 76, 77, 82, 83, 84, 86, 87, 98},
    {6, 11, 15, 16, 17, 19, 21, 22, 23, 24, 25, 26, 35, 36, 50, 52, 63, 64, 65, 66, 68, 69, 75, 82, 83, 84, 86, 102},
    {1, 6, 10, 18, 19, 25, 35, 36, 50, 51, 52, 61, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 82, 83, 84, 86, 87},
    {1, 7, 9, 10, 20, 25, 27, 28, 29, 35, 36, 59, 60, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 102},
    {1, 7, 10, 20, 25, 28, 29, 37, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 91, 92, 93, 100, 101},
    {1, 6, 10, 18, 19, 25, 28, 29, 37, 62, 67, 68, 69, 72, 73, 74, 75, 76, 77, 86, 87, 91, 92, 93, 94, 98},
    {6, 18, 19, 28, 29, 46, 55, 61, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 91, 92, 93},
    {7, 20, 28, 29, 34, 46, 55, 61, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 91, 92, 93},
    {7, 20, 28, 29, 34, 46, 55, 59, 60, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 85, 103, 104},
    {6, 18, 19, 27, 28, 29, 34, 37, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 95, 96},
    {8, 25, 28, 29, 34, 46, 53, 54, 55, 59, 60, 62, 67, 68, 69, 73, 74, 75, 85, 103, 104},
    {7, 12, 13, 14, 20, 32, 33, 44, 46, 48, 49, 53, 54, 55, 70, 71, 76, 77, 85, 97},
    {6, 18, 19, 30, 31, 32, 33, 38, 39, 42, 43, 48, 49, 72, 78, 79, 80, 81, 97},
    {7, 20, 30, 32, 33, 38, 39, 40, 43, 48, 49, 70, 71, 78, 79, 81, 88, 89, 97},
    {6, 18, 19, 32, 33, 45, 48, 49, 53, 54, 55, 70, 71, 76, 77, 85, 97},
    {6, 18, 19, 32, 33, 48, 49, 54, 55, 56, 57, 58, 70, 71, 78, 85, 97},
    {6, 18, 19, 30, 31, 38, 39, 42, 43, 47, 70, 71, 78, 79, 80, 81},
    {6, 19, 26, 32, 33, 41, 48, 49, 79, 97, 100, 101},
    {7, 20, 32, 33, 41, 70, 71, 78, 79, 90, 97},
]


# === 增强版指标收集器 ===
class MetricsCollector:
    def __init__(self):
        self.pso_start_time = None
        self.pso_end_time = None
        self.perfect_solutions_count = 0
        self.total_paths_count = 0
        self.final_fitness_scores = []
        self.pso_convergence_iterations = []
        self.pso_reset_counts = []
        self.path_execution_times = []

    def start_pso_phase(self):
        self.pso_start_time = time.time()

    def end_pso_phase(self):
        self.pso_end_time = time.time()

    def record_pso_result(self, fitness, is_perfect_match, convergence_iter=None, path_id=None, method='MI-PSO',
                          reset_count=0, execution_time=0):
        self.final_fitness_scores.append(fitness)
        self.total_paths_count += 1
        self.pso_reset_counts.append(reset_count)
        self.path_execution_times.append(execution_time)

        if is_perfect_match:
            self.perfect_solutions_count += 1
        if convergence_iter is not None:
            self.pso_convergence_iterations.append(convergence_iter)


# === 粒子类 ===
class Particle:
    def __init__(self):
        self.position = np.array([
            np.random.uniform(BOUNDS['light'][0], BOUNDS['light'][1]),
            np.random.uniform(BOUNDS['temp'][0], BOUNDS['temp'][1]),
            np.random.uniform(BOUNDS['moisture'][0], BOUNDS['moisture'][1])
        ])
        self.velocity = np.array([
            np.random.uniform(-5, 5),
            np.random.uniform(-3, 3),
            np.random.uniform(-5, 5)
        ])
        self.best_position = self.position.copy()
        self.best_fitness = 0
        self.fitness = 0


# === 单独的 MI-PSO 优化器类（已完全修复随机性重评估Bug） ===
class MIPSO:
    def __init__(self, target_path, swarm_size=20, max_iterations=3000, g2=10, th_cv=1.2):
        self.target_path = target_path
        self.swarm_size = swarm_size
        self.max_iterations = max_iterations
        self.g2 = g2
        self.th_cv = th_cv

        # 完全随机初始化
        self.particles = [Particle() for _ in range(swarm_size)]
        self.global_best_position = None
        self.global_best_fitness = 0
        self.global_best_triggered = set()  # 核心修复：记录最高分对应的真实触发路径
        self.reset_count = 0

        # 初始化适应度评估
        for particle in self.particles:
            fit, trig = self.evaluate(particle.position)
            particle.fitness = fit
            particle.best_fitness = fit
            particle.best_position = particle.position.copy()

            if fit > self.global_best_fitness:
                self.global_best_fitness = fit
                self.global_best_position = particle.position.copy()
                self.global_best_triggered = trig

    def evaluate(self, position):
        """同时返回适应度和实际触发的路径"""
        try:
            triggered = execute_Tr(position)
            if self.target_path.issubset(triggered):
                return 1.0, triggered
            intersection = len(triggered & self.target_path)
            union = len(triggered | self.target_path)
            return (intersection / union if union > 0 else 0.0), triggered
        except:
            return 0.0, set()

    def calculate_cv(self, scores):
        """实现公式 25：综合变异系数"""
        N = self.swarm_size
        mean_fit = np.mean(scores)
        if mean_fit == 0: return float('inf')

        var_term = np.sum((scores - mean_fit) ** 2) / N
        adj_diff = sum(abs(scores[2 * i + 1] - scores[2 * i]) for i in range(N // 2))
        adj_term = ((2 / N) * adj_diff) ** 2

        numerator = math.sqrt(var_term + adj_term)
        denominator = 2 * np.sum(scores)

        if denominator == 0: return float('inf')
        return numerator / denominator

    def flip_mutation(self, position, rho=0.1):
        """实现公式 27：变异反转操作"""
        new_pos = np.copy(position)
        new_pos[0] = BOUNDS['light'][0] + BOUNDS['light'][1] - position[0] + rho
        new_pos[1] = BOUNDS['temp'][0] + BOUNDS['temp'][1] - position[1] + rho
        new_pos[2] = BOUNDS['moisture'][0] + BOUNDS['moisture'][1] - position[2] + rho
        return clip_state(new_pos)

    def optimize(self):
        w = 0.7
        c1 = 1.5
        c2 = 1.5

        max_velocity = np.array([
            (BOUNDS['light'][1] - BOUNDS['light'][0]) * 0.2,
            (BOUNDS['temp'][1] - BOUNDS['temp'][0]) * 0.2,
            (BOUNDS['moisture'][1] - BOUNDS['moisture'][0]) * 0.2
        ])

        c = 0
        early_stop = False
        converged_at = self.max_iterations

        for iteration in range(self.max_iterations):
            scores = np.array([p.fitness for p in self.particles])
            c += 1

            # MI-PSO 特有的停滞检测与变异 (对应 Algorithm 4 Line 20-23)
            if c == self.g2:
                cv = self.calculate_cv(scores)
                if cv <= self.th_cv:
                    self.reset_count += 1
                    l = int(round(self.swarm_size * (1 - cv)))
                    l = min(max(l, 1), self.swarm_size)

                    selected_indices = np.random.choice(self.swarm_size, l, replace=False)
                    for idx in selected_indices:
                        new_pos = self.flip_mutation(self.particles[idx].position)
                        new_fit, new_trig = self.evaluate(new_pos)

                        if new_fit > self.particles[idx].fitness:
                            self.particles[idx].position = new_pos
                            self.particles[idx].fitness = new_fit
                            if new_fit > self.particles[idx].best_fitness:
                                self.particles[idx].best_fitness = new_fit
                                self.particles[idx].best_position = new_pos.copy()
                            if new_fit > self.global_best_fitness:
                                self.global_best_fitness = new_fit
                                self.global_best_position = new_pos.copy()
                                self.global_best_triggered = new_trig  # 更新最优路径
                c = 0

                # 标准 PSO 速度与位置更新
            for particle in self.particles:
                r1, r2 = np.random.random(3), np.random.random(3)
                particle.velocity = (w * particle.velocity +
                                     c1 * r1 * (particle.best_position - particle.position) +
                                     c2 * r2 * (self.global_best_position - particle.position))
                particle.velocity = np.clip(particle.velocity, -max_velocity, max_velocity)

                particle.position += particle.velocity
                particle.position = clip_state(particle.position)

                fit, trig = self.evaluate(particle.position)
                particle.fitness = fit

                if fit > particle.best_fitness:
                    particle.best_fitness = fit
                    particle.best_position = particle.position.copy()
                if fit > self.global_best_fitness:
                    self.global_best_fitness = fit
                    self.global_best_position = particle.position.copy()
                    self.global_best_triggered = trig  # 更新最优路径

            if self.global_best_fitness >= 1.0:
                converged_at = iteration + 1
                early_stop = True
                break

        return self.global_best_position, self.global_best_fitness, self.global_best_triggered, converged_at, early_stop, self.reset_count


# === Excel导出函数 ===
def export_multiple_runs_to_excel(all_run_results, all_run_metrics, num_runs, filename=None):
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Isolated_MIPSO_{num_runs}Runs_{timestamp}.xlsx"

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    border = Border(
        left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 工作表1: 运行汇总
    ws1 = wb.active
    ws1.title = "运行汇总"
    ws1.sheet_view.showGridLines = False

    headers = ["运行次数", "成功率", "成功数量", "平均适应度", "平均迭代次数", "运行时间(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, border, center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, (results, run_metrics) in enumerate(zip(all_run_results, all_run_metrics), start=1):
        success_count = sum(1 for r in results if r['perfect_match'])
        success_rate = (success_count / len(targetPaths)) * 100
        avg_fitness = np.mean([r['fitness'] for r in results])

        iterations_list = [r['convergence_iteration'] if r.get('convergence_iteration') is not None else 10000 for r in
                           results]
        avg_iterations = np.mean(iterations_list)
        total_time = run_metrics.pso_end_time - run_metrics.pso_start_time if run_metrics.pso_end_time else 0

        row_data = [f"运行 {run_idx}", f"{success_rate:.1f}%", f"{success_count}/{len(targetPaths)}",
                    f"{avg_fitness:.4f}", f"{avg_iterations:.1f}", f"{total_time:.2f}"]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border, cell.alignment = border, center_align
            if run_idx % 2 == 0: cell.fill = alternate_fill
            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    # 工作表2: 路径统计
    ws2 = wb.create_sheet(title="路径统计")
    ws2.sheet_view.showGridLines = False
    headers2 = ["路径编号", "成功次数", "成功率", "平均适应度", "平均迭代次数", "最小迭代", "最大迭代"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, border, center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(targetPaths)):
        success_count = sum(1 for results in all_run_results if results[path_idx]['perfect_match'])
        success_rate = (success_count / num_runs) * 100
        avg_fitness = np.mean([results[path_idx]['fitness'] for results in all_run_results])

        iterations_list = [results[path_idx]['convergence_iteration'] if results[path_idx].get(
            'convergence_iteration') is not None else 10000 for results in all_run_results]
        avg_iterations = np.mean(iterations_list)

        row_data = [f"路径 {path_idx + 1}", f"{success_count}/{num_runs}", f"{success_rate:.1f}%",
                    f"{avg_fitness:.4f}", f"{avg_iterations:.1f}", np.min(iterations_list), np.max(iterations_list)]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border, cell.alignment = border, center_align
            if (path_idx + 1) % 2 == 0: cell.fill = alternate_fill
            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    # 工作表3: 最佳粒子详情
    ws3 = wb.create_sheet(title="最佳粒子详情")
    ws3.sheet_view.showGridLines = False
    headers3 = ["路径", "运行", "最佳粒子(light,temp,moisture)", "适应度", "迭代次数", "求解方法", "生成路径"]
    col_widths3 = [10, 10, 25, 12, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, border, center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(targetPaths)):
        for run_idx, results in enumerate(all_run_results, start=1):
            result = results[path_idx]
            particle_str = f"({int(result['best_position'][0])}, {int(result['best_position'][1])}, {int(result['best_position'][2])})"
            path_str = str(sorted(list(result['triggered'])))

            convergence_iter = result['convergence_iteration'] if result.get(
                'convergence_iteration') is not None else 10000

            row_data = [f"路径{path_idx + 1}", f"运行{run_idx}", particle_str, f"{result['fitness']:.4f}",
                        convergence_iter if convergence_iter < 10000 else "-", "MI-PSO", path_str]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = left_align if col == 7 else center_align

                if result['fitness'] == 1.0:
                    cell.fill = success_fill
                elif result['fitness'] < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill
            row_idx += 1

    # 保存
    output_dir = os.path.join(os.getcwd(), "results")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"\n✓ 独立 MI-PSO 消融实验结果已导出到: {filepath}")
    return filepath


def run_single_mipso_experiment(run_num):
    print(f"\n{'=' * 50}")
    print(f"开始第 {run_num} 次运行 (独立 MI-PSO)")
    print(f"{'=' * 50}")

    run_metrics = MetricsCollector()
    run_metrics.start_pso_phase()

    pso_results = []

    for i, target_path in enumerate(targetPaths):
        path_start_time = time.time()

        mipso = MIPSO(target_path, swarm_size=20, max_iterations=3000)
        # 直接接收并使用最佳寻优过程中真实触发的 best_trig，坚决不重新 evaluate
        best_pos, best_fit, best_trig, converged_at, early_stop, reset_count = mipso.optimize()

        path_execution_time = time.time() - path_start_time

        # 使用真实的最高分数判断是否完美
        is_perfect = (best_fit >= 1.0)

        pso_results.append({
            'target_path': target_path,
            'best_position': best_pos,
            'fitness': best_fit,
            'triggered': best_trig,
            'perfect_match': is_perfect,
            'method': 'MI-PSO',
            'convergence_iteration': converged_at,
            'early_stopped': early_stop,
            'reset_count': reset_count
        })

        run_metrics.record_pso_result(
            fitness=best_fit,
            is_perfect_match=is_perfect,
            convergence_iter=converged_at if early_stop else None,
            path_id=i + 1,
            method='MI-PSO',
            reset_count=reset_count,
            execution_time=path_execution_time
        )

        status = "✓完美(MI-PSO)" if is_perfect else f"○部分({best_fit:.3f})"
        print(f"  路径{i + 1}: {status} | 变异触发次数: {reset_count} | 耗时 {path_execution_time:.2f}s")

    run_metrics.end_pso_phase()

    success_count = sum(1 for r in pso_results if r['perfect_match'])
    success_rate = (success_count / len(targetPaths)) * 100
    pso_time = run_metrics.pso_end_time - run_metrics.pso_start_time

    print(f"\n第{run_num}次运行完成: 成功率 {success_rate:.1f}% | 耗时 {pso_time:.2f}秒")
    return pso_results, run_metrics


def run_multiple_mipso_experiments(num_runs):
    all_run_results = []
    all_run_metrics = []
    total_start = time.time()

    for run_num in range(1, num_runs + 1):
        results, metrics = run_single_mipso_experiment(run_num)
        all_run_results.append(results)
        all_run_metrics.append(metrics)

    total_time = time.time() - total_start
    print(f"\n{'=' * 50}\n全部{num_runs}次独立 MI-PSO 运行完成! 总耗时: {total_time:.2f}秒\n{'=' * 50}")

    return all_run_results, all_run_metrics


if __name__ == "__main__":
    if len(sys.argv) > 1:
        try:
            NUM_RUNS = int(sys.argv[1])
        except ValueError:
            pass

    print("=" * 70)
    print(" 独立 MI-PSO 消融实验 (无 DQN 介入，完美修复随机性 Bug)")
    print("=" * 70)

    all_run_results, all_run_metrics = run_multiple_mipso_experiments(NUM_RUNS)
    export_multiple_runs_to_excel(all_run_results, all_run_metrics, NUM_RUNS)