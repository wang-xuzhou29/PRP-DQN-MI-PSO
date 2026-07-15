import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import math

def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

def execute_validation_rules(x, y, z):
    b = {}  # 使用字典存储触发的规则

    if (30 < x < 170) != (30 < x < 100): b[0] = 1
    if (30 < x < 170) != (30 < x < 190): b[1] = 2
    if (30 < x < 170) != (30 < x < 140): b[2] = 3
    if (30 < x < 170) != (30 < x < 150): b[3] = 4
    if (30 < x < 170) != (30 < x < 160): b[4] = 5
    if (100 < y < 200) != (150 < y < 200): b[5] = 6
    if (100 < y < 200) != (50 < y < 200): b[6] = 7
    if (100 < y < 200) != (100 < y < 150): b[7] = 8
    if (1 < z < 4) != (1 < z < 3): b[8] = 9
    if (x > 10 and x < 152) != (x > 10 and x < 100): b[9] = 10

    # 规则11-20
    if (x > 40 and x < 152) != (x > 40 and x < 202): b[10] = 11
    if (abs(x - 50) < 4) != (abs(x - 50) < 2): b[11] = 12
    if (abs(x - 50) < 4) != (abs(x + 50) < 4): b[12] = 13
    if (abs(x - 50) < 4) != (abs(x * 50) < 4): b[13] = 14
    if (x > 135) != (x > 215): b[14] = 15
    if (x > 135) != (x > 235): b[15] = 16
    if (x > 135) != (x > 225): b[16] = 17
    if (y > 100) != (y > 120): b[17] = 18
    if (y > 100) != (y > 140): b[18] = 19
    if (y > 100) != (y > 50): b[19] = 20

    # 规则21-30
    if (x < 165) != (x < 265): b[20] = 21
    if (x < 165) != (x < 245): b[21] = 22
    if (x < 165) != (x < 262): b[22] = 23
    if (y < 150) != (y < 130): b[23] = 24
    if (y < 150) != (x < 100): b[24] = 25
    if (y < 190) != (y < 120): b[25] = 26
    if ((x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 125) != (
            (x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 115): b[26] = 27
    if ((x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 125) != (
            (x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 85): b[27] = 28
    if ((x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 125) != (
            (x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 25): b[28] = 29
    if (x / (y / 30) > 1 and x / (y / 30) < 2) != (x / (y / 30) > 1 and x / (y / 30) < 4): b[29] = 30

    # 规则31-40
    if (x / (y / 30) > 1 and x / (y / 30) < 2) != (x / (y / 30) > 1 and x / (y / 30) < 3): b[30] = 31
    if (x / (z / 2) > 2 and x / (z / 2) < 4) != (x / (z / 2) > 1 and x / (z / 2) < 4): b[31] = 32
    if (x / (z / 2) > 2 and x / (z / 2) < 4) != (x / (z / 3) > 2 and x / (z / 2) < 4): b[32] = 33
    if (x / (z / 2) > 2 and x / (z / 2) < 4) != (x / (z / 2) > 2 and x / (z / 3) < 4): b[33] = 34
    if ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 55) != ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 155): b[
        34] = 35
    if ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 55) != ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 125): b[
        35] = 36
    if ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 55) != ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 25): b[
        36] = 37
    if (x - y / 30 > -10 and x - y / 30 < 10) != (x - y / 30 > -10 and x - y / 30 < 5): b[37] = 38
    if (x - y / 30 > -10 and x - y / 30 < 10) != (x - y / 30 > -10 and x - y / 30 < 6): b[38] = 39
    if (x - y / 30 > -10 and x - y / 30 < 10) != (x - y / 30 > -10 and x - y / 30 < 8): b[39] = 40

    # 规则41-50
    if (abs(x / (y / 30) - 1.0) < 1) != (abs(x / (y / 30) + 1.0) < 1): b[40] = 41
    if (abs(x / (y / 30) - 1.0) < 1) != (abs(x / (y / 30) - 1.0) < 2): b[41] = 42
    if (abs(x / (y / 30) - 1.0) < 1) != (abs(x / (y / 30) - 1.0) < 3): b[42] = 43
    if (x / ((y / 30) + 20) > 1 and x / ((y / 30) + 20) < 2) != (x / ((y / 30) + 20) > 1 and x / ((y / 20) + 20) < 2):
        b[43] = 44
    if (x / ((y / 30) + 20) > 1 and x / ((y / 30) + 20) < 2) != (x / ((y / 30) + 20) > 1 and x / ((y / 40) + 20) < 2):
        b[44] = 45
    if (x / ((y / 30) + 20) > 1 and x / ((y / 30) + 20) < 2) != (x / ((y / 30) + 20) > 1 and x / ((y / 30) + 40) < 2):
        b[45] = 46
    if ((z / 50) / (x / 50) > 1 and (z / 50) / (x / 50) < 3) != ((z / 50) / (x / 50) > 1 and (z / 50) / (x / 50) < 4):
        b[46] = 47
    if ((z / 50) / (x / 50) > 1 and (z / 50) / (x / 50) < 3) != ((z / 50) / (x / 50) > 2 and (z / 50) / (x / 50) < 3):
        b[47] = 48
    if ((z / 50) / (x / 50) > 1 and (z / 50) / (x / 50) < 3) != ((z / 50) / (x / 50) > 3 and (z / 50) / (x / 50) < 3):
        b[48] = 49
    if ((x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 45) != (
            (x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 70): b[49] = 50

    # 规则51-60
    if ((x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 45) != (
            (x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 60): b[50] = 51
    if ((x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 45) != (
            (x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 6 < 45): b[51] = 52
    if (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 8) != (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 48): b[
        52] = 53
    if (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 8) != (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 50): b[
        53] = 54
    if (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 8) != (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 58): b[
        54] = 55
    if ((x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 200) != (
            (x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 400): b[55] = 56
    if ((x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 200) != (
            (x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 300): b[56] = 57
    if ((x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 200) != (
            (x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 500): b[57] = 58
    if (abs((x + y / 30) - 100) < 15) != (abs((x + y / 30) - 100) < 35): b[58] = 59
    if (abs((x + y / 30) - 100) < 15) != (abs((x + y / 30) - 100) < 45): b[59] = 60

    # 规则61-70
    if (abs((x + y / 30) - 100) < 15) != (abs((x + y / 30) - 100) < 10): b[60] = 61
    if (x / 50 > 1 and x / 50 < 4) != (x / 50 > 3 and x / 50 < 4): b[61] = 62
    if (x / 50 > 1 and x / 50 < 4) != (x / 50 > 1 and x / 50 < 5): b[62] = 63
    if (x < 135) != (x < 215): b[63] = 64
    if (x < 135) != (x < 225): b[64] = 65
    if (x < 135) != (x < 225): b[65] = 66
    if (x > 65) != (x > 165): b[66] = 67
    if (x > 65) != (x > 215): b[67] = 68
    if (x > 65) != (x > 251): b[68] = 69
    if (y < 110) != (y < 80): b[69] = 70

    # 规则71-80
    if (y < 110) != (y < 50): b[70] = 71
    if (y < 110) != (y < 119): b[71] = 72
    if (x < 32 or x > 68) != (x < 32 or x > 168): b[72] = 73
    if (x < 32 or x > 68) != (x < 32 or x > 118): b[73] = 74
    if (x < 32 or x > 68) != (x < 32 or x > 218): b[74] = 75
    if (x < 35 and y < 120) != (x < 135 and y < 120): b[75] = 76
    if (x < 35 and y < 120) != (x < 235 and y < 120): b[76] = 77
    if (x < 35 and y < 120) != (x < 35 and y < 60): b[77] = 78
    if (abs(x / (y / 30) - 1.0) > 1) != (abs(x / (y / 30) - 3.0) > 1): b[78] = 79
    if (abs(x / (y / 30) - 1.0) > 1) != (abs(x / (y / 30) - 1.0) > 2): b[79] = 80

    # 规则81-90
    if (abs(x / (y / 30) - 1.0) > 1) != (abs(x / (y / 30) - 1.0) > 3): b[80] = 81
    if ((x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 135) != (
            (x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 235): b[81] = 82
    if ((x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 135) != (
            (x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 335): b[82] = 83
    if ((x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 135) != (
            (x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 435): b[83] = 84
    if (x < 30 or x > 70) != (x < 30 or x * 7 > 70): b[84] = 85
    if (x < 30 or x > 70) != (x < 30 or x > 270): b[85] = 86
    if (x < 30 or x > 70) != (x < 30 or x > 170): b[86] = 87
    if (x / (y / 30) > 1 and x / (y / 30) < 3) != (x / (y / 30) > 1 and x / (y / 20) < 3): b[87] = 88
    if (x / (y / 30) > 1 and x / (y / 30) < 3) != (x / (y / 30) > 1 and x / (y / 30) < 4): b[88] = 89
    if (x / (y / 30) > 1 and x / (y / 30) < 3) != (x / (y / 20) > 1 and x / (y / 30) < 3): b[89] = 90

    # 规则91-100
    if ((x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 112) != (
            (x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 62): b[90] = 91
    if ((x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 112) != (
            (x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 92): b[91] = 92
    if ((x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 112) != (
            (x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 82): b[92] = 93
    if (abs((x + y / 30) / 2 - 50) < 1) != (abs((x + y / 30) / 2 - 50) < 4): b[93] = 94
    if (abs((x + y / 30) / 2 - 50) < 1) != (abs((x + y / 30) / 2 + 50) < 1): b[94] = 95
    if (abs((x + y / 30) / 2 - 50) < 1) != (abs((x + y / 20) / 2 + 50) < 1): b[95] = 96
    if (x / (z / 2) > 2 and x / (z / 2) < 4) != (x / (z / 2) > 1 and x / (z / 2) < 4): b[96] = 97
    if (x / (z / 2) > 20 and x / (z / 2) < 100) != (x / (z / 3) > 20 and x / (z / 2) < 100): b[97] = 98
    if (x / (z / 2) > 20 and x / (z / 2) < 100) != (x / (z / 2) > 20 and x / (z / 2) < 150): b[98] = 99
    if ((y / 30) / (z / 50) > 38 and (y / 30) / (z / 50) < 42) != (
            (y / 30) / (z / 50) > 48 and (y / 30) / (z / 50) < 42): b[99] = 100

    # 规则101-110
    if ((y / 30) / (z / 50) > 38 and (y / 30) / (z / 50) < 42) != (
            (y / 30) / (z / 50) > 58 and (y / 30) / (z / 50) < 42): b[100] = 101
    if ((y / 30) / (z / 50) > 38 and (y / 30) / (z / 50) < 42) != (
            (y / 30) / (z / 50) > 38 and (y / 30) / (z / 50) < 72): b[101] = 102
    if ((x - 40) / 20 > 1 and (x - 40) / 20 < 6) != ((x - 40) / 30 > 1 and (x - 40) / 20 < 7): b[102] = 103
    if ((x - 40) / 20 > 1 and (x - 40) / 20 < 6) != ((x - 50) / 20 > 1 and (x - 40) / 20 < 7): b[103] = 104
    if ((x - 40) / 20 > 1 and (x - 40) / 20 < 6) != ((x - 40) / 20 > 1 and (x - 50) / 20 < 7): b[104] = 105

    triggered = set(b.values())
    return triggered

def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """"""
    generated_path = execute_validation_rules(particle[0], particle[1], particle[2])

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # : x:1-50, y:1-50, z:1-50
        self.bounds = bounds if bounds else [(1, 100), (1, 100), (1, 100)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """target pathsPSO"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(particles[i][0], particles[i][1], particles[i][2]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(gbest_particle[0], gbest_particle[1], gbest_particle[2]),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """Path PSO"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - Path ")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations")
    print(f"Path : {len(target_paths)}")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"Path {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "" if result['success'] else f"({result['best_fitness']:.3f})"
        print(f"{status} | {result['time']:.2f}s | iterations{result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f": {success_count}/{len(target_paths)} ({success_rate:.1f}%) | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs}")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations, {len(target_paths)}Path ")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"---  {run_idx}/{num_runs} ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f": {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs} runcompleted | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 1: 
    ws1 = wb.active
    ws1.title = ""
    ws1.sheet_view.showGridLines = False

    headers = ["Run", "", "", "", "Average Iterations", "(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f" {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # 2: Path 
    ws2 = wb.create_sheet(title="Path ")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Path ID", "", "", "", "Average Iterations", "Minimum Iterations", "Maximum Iterations"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"Path  {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # 3: 
    ws3 = wb.create_sheet(title="")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Path ", "", "(x,y,z)", "", "Iterations", "Path "]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"{run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # 4: target paths
    ws4 = wb.create_sheet(title="target paths")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Path ID", "target paths", ""]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"Path  {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f" : {filename}")
    print(f"{'=' * 70}")
    print(f":")
    print(f"  1.        - {len(all_results)} run")
    print(f"  2. Path        - Path ")
    print(f"  3.    -  runPath ")
    print(f"  4. target paths       - target paths")
    print(f"{'=' * 70}\n")

    return filename


def main():


    target_paths = [
        {1, 3, 4, 5, 6, 9, 11, 15, 16, 17, 18, 19, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 67, 68, 69, 73, 75,
         77,
         82, 83, 84, 86, 87, 99, 102, 103, 104, 105},
        {1, 3, 4, 5, 6, 9, 11, 15, 16, 17, 18, 19, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 70, 71, 75,
         77,
         82, 83, 84, 86, 87, 99, 102, 103, 104, 105},
        {1, 3, 4, 5, 6, 11, 15, 16, 17, 18, 19, 21, 22, 23, 25, 35, 36, 50, 52, 64, 65, 66, 68, 69, 70, 71, 73, 75, 77,
         82,
         83, 84, 86, 87, 98, 103, 104, 105},
        {1, 3, 4, 5, 6, 11, 15, 16, 17, 19, 21, 22, 23, 24, 25, 26, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 73, 75, 82,
         83,
         84, 86, 87, 102, 103, 104, 105},
        {1, 3, 4, 5, 7, 11, 15, 16, 17, 20, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 71, 73, 75, 77, 82,
         83,
         84, 86, 87, 98, 103, 104, 105},
        {2, 6, 9, 11, 15, 16, 17, 18, 19, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 70, 71, 75, 77, 82,
         83,
         84, 86, 99, 102, 103, 104, 105},
        {2, 7, 11, 15, 16, 17, 20, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 70, 71, 75, 77, 82, 83, 84,
         86,
         87, 100, 101, 103, 104, 105},
        {1, 3, 4, 5, 6, 11, 15, 16, 17, 18, 19, 25, 35, 36, 50, 51, 52, 64, 65, 66, 67, 68, 69, 72, 73, 75, 77, 82, 83,
         84,
         86, 87, 98, 103, 104, 105},
        {1, 3, 4, 5, 8, 9, 11, 15, 16, 17, 21, 22, 23, 26, 35, 36, 50, 51, 52, 64, 65, 66, 67, 68, 69, 73, 75, 82, 83,
         84,
         86, 87, 99, 103, 104, 105},
        {1, 3, 7, 9, 10, 15, 16, 17, 20, 25, 35, 36, 50, 51, 52, 60, 62, 64, 65, 66, 67, 68, 69, 70, 71, 73, 75, 77, 82,
         83,
         84, 86, 87, 102},
        {1, 6, 10, 18, 19, 25, 35, 36, 50, 51, 52, 59, 60, 62, 67, 68, 69, 70, 71, 73, 75, 76, 77, 82, 83, 84, 86, 87,
         98},
        {6, 11, 15, 16, 17, 19, 21, 22, 23, 24, 25, 26, 35, 36, 50, 52, 63, 64, 65, 66, 68, 69, 75, 82, 83, 84, 86,
         102},
        {1, 6, 10, 18, 19, 25, 35, 36, 50, 51, 52, 61, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 82, 83, 84, 86, 87},
        {1, 7, 9, 10, 20, 25, 27, 28, 29, 35, 36, 59, 60, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 102},
        {1, 7, 10, 20, 25, 28, 29, 37, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 91, 92, 93, 100, 101},
        {1, 6, 10, 18, 19, 25, 28, 29, 37, 62, 67, 68, 69, 72, 73, 74, 75, 76, 77, 86, 87, 91, 92, 93, 94, 98},
        {6, 18, 19, 28, 29, 46, 55, 61, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 91, 92, 93},
        {7, 20, 28, 29, 34, 46, 55, 61, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 91, 92, 93},
        {7, 20, 28, 29, 34, 46, 55, 59, 60, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 85, 103, 104},
        {6, 18, 19, 27, 28, 29, 34, 37, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 95, 96},
        {8, 25, 28, 29, 34, 46, 53, 54, 55, 59, 60, 62, 67, 68, 69, 73, 74, 75, 85, 103, 104},
        {7, 12, 13, 14, 20, 32, 33, 44, 46, 48, 49, 53, 54, 55, 70, 71, 76, 77, 85, 97},
        {6, 18, 19, 30, 31, 32, 33, 38, 39, 42, 43, 48, 49, 72, 78, 79, 80, 81, 97},
        {7, 20, 30, 32, 33, 38, 39, 40, 43, 48, 49, 70, 71, 78, 79, 81, 88, 89, 97},
        {6, 18, 19, 32, 33, 45, 48, 49, 53, 54, 55, 70, 71, 76, 77, 85, 97},
        {6, 18, 19, 32, 33, 48, 49, 54, 55, 56, 57, 58, 70, 71, 78, 85, 97},
        {6, 18, 19, 30, 31, 38, 39, 42, 43, 47, 70, 71, 78, 79, 80, 81},
        {6, 19, 26, 32, 33, 41, 48, 49, 79, 97, 100, 101},
        {7, 20, 32, 33, 41, 70, 71, 78, 79, 90, 97},
    ]

    print("=" * 70)
    print("baseline PSO")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("Program completed")


if __name__ == "__main__":
    main()