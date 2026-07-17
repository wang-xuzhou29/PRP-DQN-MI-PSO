import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import math  # 添加 math 导入，供 section2 使用

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === 全局取值范围（按要求设置为 X:1~300, Y:1~200, Z:1~5） ===
MIN_X = 1
MAX_X = 300
MIN_Y = 1
MAX_Y = 200
MIN_Z = 1
MAX_Z = 5

# === 归一化/反归一化 ===
def normalize_state(state):
    """将状态归一化到 [0, 1] 区间"""
    weather_norm = (state[0] - MIN_X) / (MAX_X - MIN_X)
    time_norm = (state[1] - MIN_Y) / (MAX_Y - MIN_Y)
    z_norm = (state[2] - MIN_Z) / (MAX_Z - MIN_Z)
    return (weather_norm, time_norm, z_norm)

def denormalize_state(state_norm):
    """将归一化状态还原"""
    weather = int(round(state_norm[0] * (MAX_X - MIN_X) + MIN_X))
    time_period = int(round(state_norm[1] * (MAX_Y - MIN_Y) + MIN_Y))
    z = int(round(state_norm[2] * (MAX_Z - MIN_Z) + MIN_Z))

    # 边界保护
    weather = np.clip(weather, MIN_X, MAX_X)
    time_period = np.clip(time_period, MIN_Y, MAX_Y)
    z = np.clip(z, MIN_Z, MAX_Z)

    return (weather, time_period, z)

def normalize_value(value, min_val, max_val):
    return (value - min_val) / (max_val - min_val)

def denormalize_value(value_norm, min_val, max_val):
    return int(round(value_norm * (max_val - min_val) + min_val))

# === 安全除法 ===
def safe_divide(numerator, denominator, default=0.0):
    if denominator == 0:
        return default
    return numerator / denominator

# === reward function ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10

    if target_path.issubset(triggered):
        reward += 1

    if prev_triggered is not None:
        prev_sim = jaccard_similarity(prev_triggered, target_path)
        improvement = sim - prev_sim
        reward += improvement * 5

    return reward


# ========== 规则触发函数（section2_flow_pressure_density_hybrid） ==========
def section2_flow_pressure_density_hybrid(x, y, z):
    b = {}  # 使用字典存储触发的规则

    if (30 < x < 170) != (30 < x < 100): b[0] = 1
    if (30 < x < 170) != (30 < x < 190): b[1] = 2
    if (30 < x < 170) != (30 < x < 140): b[2] = 3
    if (30 < x < 170) != (30 < x < 150): b[3] = 4
    if (30 < x < 170) != (30 < x < 160): b[4] = 5
    if (100 < y < 200) != (150 < y < 200): b[5] = 6
    if (100 < y < 200) != (50 < y < 200): b[6] = 7
    if (100 < y < 200) != (100 < y < 150): b[7] = 8
    if (1 < z < 4) != (1 < z < 3): b[8] = 9
    if (x > 10 and x < 152) != (x > 10 and x < 100): b[9] = 10

    # 规则11-20
    if (x > 40 and x < 152) != (x > 40 and x < 202): b[10] = 11
    if (abs(x - 50) < 4) != (abs(x - 50) < 2): b[11] = 12
    if (abs(x - 50) < 4) != (abs(x + 50) < 4): b[12] = 13
    if (abs(x - 50) < 4) != (abs(x * 50) < 4): b[13] = 14
    if (x > 135) != (x > 215): b[14] = 15
    if (x > 135) != (x > 235): b[15] = 16
    if (x > 135) != (x > 225): b[16] = 17
    if (y > 100) != (y > 120): b[17] = 18
    if (y > 100) != (y > 140): b[18] = 19
    if (y > 100) != (y > 50): b[19] = 20

    # 规则21-30
    if (x < 165) != (x < 265): b[20] = 21
    if (x < 165) != (x < 245): b[21] = 22
    if (x < 165) != (x < 262): b[22] = 23
    if (y < 150) != (y < 130): b[23] = 24
    if (y < 150) != (x < 100): b[24] = 25
    if (y < 190) != (y < 120): b[25] = 26
    if ((x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 125) != (
            (x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 115): b[26] = 27
    if ((x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 125) != (
            (x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 85): b[27] = 28
    if ((x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 125) != (
            (x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 25): b[28] = 29
    if (x / (y / 30) > 1 and x / (y / 30) < 2) != (x / (y / 30) > 1 and x / (y / 30) < 4): b[29] = 30

    # 规则31-40
    if (x / (y / 30) > 1 and x / (y / 30) < 2) != (x / (y / 30) > 1 and x / (y / 30) < 3): b[30] = 31
    if (x / (z / 2) > 2 and x / (z / 2) < 4) != (x / (z / 2) > 1 and x / (z / 2) < 4): b[31] = 32
    if (x / (z / 2) > 2 and x / (z / 2) < 4) != (x / (z / 3) > 2 and x / (z / 2) < 4): b[32] = 33
    if (x / (z / 2) > 2 and x / (z / 2) < 4) != (x / (z / 2) > 2 and x / (z / 3) < 4): b[33] = 34
    if ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 55) != ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 155): b[
        34] = 35
    if ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 55) != ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 125): b[
        35] = 36
    if ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 55) != ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 25): b[
        36] = 37
    if (x - y / 30 > -10 and x - y / 30 < 10) != (x - y / 30 > -10 and x - y / 30 < 5): b[37] = 38
    if (x - y / 30 > -10 and x - y / 30 < 10) != (x - y / 30 > -10 and x - y / 30 < 6): b[38] = 39
    if (x - y / 30 > -10 and x - y / 30 < 10) != (x - y / 30 > -10 and x - y / 30 < 8): b[39] = 40

    # 规则41-50
    if (abs(x / (y / 30) - 1.0) < 1) != (abs(x / (y / 30) + 1.0) < 1): b[40] = 41
    if (abs(x / (y / 30) - 1.0) < 1) != (abs(x / (y / 30) - 1.0) < 2): b[41] = 42
    if (abs(x / (y / 30) - 1.0) < 1) != (abs(x / (y / 30) - 1.0) < 3): b[42] = 43
    if (x / ((y / 30) + 20) > 1 and x / ((y / 30) + 20) < 2) != (x / ((y / 30) + 20) > 1 and x / ((y / 20) + 20) < 2):
        b[43] = 44
    if (x / ((y / 30) + 20) > 1 and x / ((y / 30) + 20) < 2) != (x / ((y / 30) + 20) > 1 and x / ((y / 40) + 20) < 2):
        b[44] = 45
    if (x / ((y / 30) + 20) > 1 and x / ((y / 30) + 20) < 2) != (x / ((y / 30) + 20) > 1 and x / ((y / 30) + 40) < 2):
        b[45] = 46
    if ((z / 50) / (x / 50) > 1 and (z / 50) / (x / 50) < 3) != ((z / 50) / (x / 50) > 1 and (z / 50) / (x / 50) < 4):
        b[46] = 47
    if ((z / 50) / (x / 50) > 1 and (z / 50) / (x / 50) < 3) != ((z / 50) / (x / 50) > 2 and (z / 50) / (x / 50) < 3):
        b[47] = 48
    if ((z / 50) / (x / 50) > 1 and (z / 50) / (x / 50) < 3) != ((z / 50) / (x / 50) > 3 and (z / 50) / (x / 50) < 3):
        b[48] = 49
    if ((x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 45) != (
            (x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 70): b[49] = 50

    # 规则51-60
    if ((x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 45) != (
            (x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 60): b[50] = 51
    if ((x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 45) != (
            (x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 6 < 45): b[51] = 52
    if (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 8) != (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 48): b[
        52] = 53
    if (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 8) != (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 50): b[
        53] = 54
    if (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 8) != (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 58): b[
        54] = 55
    if ((x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 200) != (
            (x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 400): b[55] = 56
    if ((x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 200) != (
            (x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 300): b[56] = 57
    if ((x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 200) != (
            (x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 500): b[57] = 58
    if (abs((x + y / 30) - 100) < 15) != (abs((x + y / 30) - 100) < 35): b[58] = 59
    if (abs((x + y / 30) - 100) < 15) != (abs((x + y / 30) - 100) < 45): b[59] = 60

    # 规则61-70
    if (abs((x + y / 30) - 100) < 15) != (abs((x + y / 30) - 100) < 10): b[60] = 61
    if (x / 50 > 1 and x / 50 < 4) != (x / 50 > 3 and x / 50 < 4): b[61] = 62
    if (x / 50 > 1 and x / 50 < 4) != (x / 50 > 1 and x / 50 < 5): b[62] = 63
    if (x < 135) != (x < 215): b[63] = 64
    if (x < 135) != (x < 225): b[64] = 65
    if (x < 135) != (x < 225): b[65] = 66
    if (x > 65) != (x > 165): b[66] = 67
    if (x > 65) != (x > 215): b[67] = 68
    if (x > 65) != (x > 251): b[68] = 69
    if (y < 110) != (y < 80): b[69] = 70

    # 规则71-80
    if (y < 110) != (y < 50): b[70] = 71
    if (y < 110) != (y < 119): b[71] = 72
    if (x < 32 or x > 68) != (x < 32 or x > 168): b[72] = 73
    if (x < 32 or x > 68) != (x < 32 or x > 118): b[73] = 74
    if (x < 32 or x > 68) != (x < 32 or x > 218): b[74] = 75
    if (x < 35 and y < 120) != (x < 135 and y < 120): b[75] = 76
    if (x < 35 and y < 120) != (x < 235 and y < 120): b[76] = 77
    if (x < 35 and y < 120) != (x < 35 and y < 60): b[77] = 78
    if (abs(x / (y / 30) - 1.0) > 1) != (abs(x / (y / 30) - 3.0) > 1): b[78] = 79
    if (abs(x / (y / 30) - 1.0) > 1) != (abs(x / (y / 30) - 1.0) > 2): b[79] = 80

    # 规则81-90
    if (abs(x / (y / 30) - 1.0) > 1) != (abs(x / (y / 30) - 1.0) > 3): b[80] = 81
    if ((x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 135) != (
            (x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 235): b[81] = 82
    if ((x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 135) != (
            (x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 335): b[82] = 83
    if ((x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 135) != (
            (x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 435): b[83] = 84
    if (x < 30 or x > 70) != (x < 30 or x * 7 > 70): b[84] = 85
    if (x < 30 or x > 70) != (x < 30 or x > 270): b[85] = 86
    if (x < 30 or x > 70) != (x < 30 or x > 170): b[86] = 87
    if (x / (y / 30) > 1 and x / (y / 30) < 3) != (x / (y / 30) > 1 and x / (y / 20) < 3): b[87] = 88
    if (x / (y / 30) > 1 and x / (y / 30) < 3) != (x / (y / 30) > 1 and x / (y / 30) < 4): b[88] = 89
    if (x / (y / 30) > 1 and x / (y / 30) < 3) != (x / (y / 20) > 1 and x / (y / 30) < 3): b[89] = 90

    # 规则91-100
    if ((x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 112) != (
            (x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 62): b[90] = 91
    if ((x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 112) != (
            (x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 92): b[91] = 92
    if ((x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 112) != (
            (x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 82): b[92] = 93
    if (abs((x + y / 30) / 2 - 50) < 1) != (abs((x + y / 30) / 2 - 50) < 4): b[93] = 94
    if (abs((x + y / 30) / 2 - 50) < 1) != (abs((x + y / 30) / 2 + 50) < 1): b[94] = 95
    if (abs((x + y / 30) / 2 - 50) < 1) != (abs((x + y / 20) / 2 + 50) < 1): b[95] = 96
    if (x / (z / 2) > 2 and x / (z / 2) < 4) != (x / (z / 2) > 1 and x / (z / 2) < 4): b[96] = 97
    if (x / (z / 2) > 20 and x / (z / 2) < 100) != (x / (z / 3) > 20 and x / (z / 2) < 100): b[97] = 98
    if (x / (z / 2) > 20 and x / (z / 2) < 100) != (x / (z / 2) > 20 and x / (z / 2) < 150): b[98] = 99
    if ((y / 30) / (z / 50) > 38 and (y / 30) / (z / 50) < 42) != (
            (y / 30) / (z / 50) > 48 and (y / 30) / (z / 50) < 42): b[99] = 100

    # 规则101-110
    if ((y / 30) / (z / 50) > 38 and (y / 30) / (z / 50) < 42) != (
            (y / 30) / (z / 50) > 58 and (y / 30) / (z / 50) < 42): b[100] = 101
    if ((y / 30) / (z / 50) > 38 and (y / 30) / (z / 50) < 42) != (
            (y / 30) / (z / 50) > 38 and (y / 30) / (z / 50) < 72): b[101] = 102
    if ((x - 40) / 20 > 1 and (x - 40) / 20 < 6) != ((x - 40) / 30 > 1 and (x - 40) / 20 < 7): b[102] = 103
    if ((x - 40) / 20 > 1 and (x - 40) / 20 < 6) != ((x - 50) / 20 > 1 and (x - 40) / 20 < 7): b[103] = 104
    if ((x - 40) / 20 > 1 and (x - 40) / 20 < 6) != ((x - 40) / 20 > 1 and (x - 50) / 20 < 7): b[104] = 105

    triggered = set(b.values())
    return triggered

# === 目标路径组 ===
targetPaths = [
    {1, 3, 4, 5, 6, 9, 11, 15, 16, 17, 18, 19, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 67, 68, 69, 73, 75, 77,
     82, 83, 84, 86, 87, 99, 102, 103, 104, 105},
    {1, 3, 4, 5, 6, 9, 11, 15, 16, 17, 18, 19, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 70, 71, 75, 77,
     82, 83, 84, 86, 87, 99, 102, 103, 104, 105},
    {1, 3, 4, 5, 6, 11, 15, 16, 17, 18, 19, 21, 22, 23, 25, 35, 36, 50, 52, 64, 65, 66, 68, 69, 70, 71, 73, 75, 77, 82,
     83, 84, 86, 87, 98, 103, 104, 105},
    {1, 3, 4, 5, 6, 11, 15, 16, 17, 19, 21, 22, 23, 24, 25, 26, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 73, 75, 82, 83,
     84, 86, 87, 102, 103, 104, 105},
    {1, 3, 4, 5, 7, 11, 15, 16, 17, 20, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 71, 73, 75, 77, 82, 83,
     84, 86, 87, 98, 103, 104, 105},
    {2, 6, 9, 11, 15, 16, 17, 18, 19, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 70, 71, 75, 77, 82, 83,
     84, 86, 99, 102, 103, 104, 105},
    {2, 7, 11, 15, 16, 17, 20, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 70, 71, 75, 77, 82, 83, 84, 86,
     87, 100, 101, 103, 104, 105},
    {1, 3, 4, 5, 6, 11, 15, 16, 17, 18, 19, 25, 35, 36, 50, 51, 52, 64, 65, 66, 67, 68, 69, 72, 73, 75, 77, 82, 83, 84,
     86, 87, 98, 103, 104, 105},
    {1, 3, 4, 5, 8, 9, 11, 15, 16, 17, 21, 22, 23, 26, 35, 36, 50, 51, 52, 64, 65, 66, 67, 68, 69, 73, 75, 82, 83, 84,
     86, 87, 99, 103, 104, 105},
    {1, 3, 7, 9, 10, 15, 16, 17, 20, 25, 35, 36, 50, 51, 52, 60, 62, 64, 65, 66, 67, 68, 69, 70, 71, 73, 75, 77, 82, 83,
     84, 86, 87, 102},
    {1, 6, 10, 18, 19, 25, 35, 36, 50, 51, 52, 59, 60, 62, 67, 68, 69, 70, 71, 73, 75, 76, 77, 82, 83, 84, 86, 87, 98},
    {6, 11, 15, 16, 17, 19, 21, 22, 23, 24, 25, 26, 35, 36, 50, 52, 63, 64, 65, 66, 68, 69, 75, 82, 83, 84, 86, 102},
    {1, 6, 10, 18, 19, 25, 35, 36, 50, 51, 52, 61, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 82, 83, 84, 86, 87},
    {1, 7, 9, 10, 20, 25, 27, 28, 29, 35, 36, 59, 60, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 102},
    {1, 7, 10, 20, 25, 28, 29, 37, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 91, 92, 93, 100, 101},
    {1, 6, 10, 18, 19, 25, 28, 29, 37, 62, 67, 68, 69, 72, 73, 74, 75, 76, 77, 86, 87, 91, 92, 93, 94, 98},
    {6, 18, 19, 28, 29, 46, 55, 61, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 91, 92, 93},
    {7, 20, 28, 29, 34, 46, 55, 61, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 91, 92, 93},
    {7, 20, 28, 29, 34, 46, 55, 59, 60, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 85, 103, 104},
    {6, 18, 19, 27, 28, 29, 34, 37, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 95, 96},
    {8, 25, 28, 29, 34, 46, 53, 54, 55, 59, 60, 62, 67, 68, 69, 73, 74, 75, 85, 103, 104},
    {7, 12, 13, 14, 20, 32, 33, 44, 46, 48, 49, 53, 54, 55, 70, 71, 76, 77, 85, 97},
    {6, 18, 19, 30, 31, 32, 33, 38, 39, 42, 43, 48, 49, 72, 78, 79, 80, 81, 97},
    {7, 20, 30, 32, 33, 38, 39, 40, 43, 48, 49, 70, 71, 78, 79, 81, 88, 89, 97},
    {6, 18, 19, 32, 33, 45, 48, 49, 53, 54, 55, 70, 71, 76, 77, 85, 97},
    {6, 18, 19, 32, 33, 48, 49, 54, 55, 56, 57, 58, 70, 71, 78, 85, 97},
    {6, 18, 19, 30, 31, 38, 39, 42, 43, 47, 70, 71, 78, 79, 80, 81},
    {6, 19, 26, 32, 33, 41, 48, 49, 79, 97, 100, 101},
    {7, 20, 32, 33, 41, 70, 71, 78, 79, 90, 97},
]

# 将路径列表转换为集合列表
target_paths = [set(path) for path in targetPaths]
NUM_PATHS = len(target_paths)

def jaccard_similarity(set1, set2):
    """Compute Jaccard similarity"""
    if not set2:
        return 0.0

    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0

# === Path Similarity ===
def compute_path_similarity_matrix(paths):
    """Path Similarity"""
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix

def group_paths_by_similarity(paths, threshold_percentile=50):
    """SimilarityPath """
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.percentile(avg_sim_scores, threshold_percentile)

    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]

    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)

    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group

# === Sample generation ===
def compute_robustness(state, path, sample_size=9):
    """()"""
    base = section2_flow_pressure_density_hybrid(state[0], state[1], state[2])
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0

    # :
    # weather1-6, +/-1
    # time_period1-6, +/-1
    # z1-60, +/-6(10%), +/-3(5%)
    deltas = [
        (-1, -1, -6), (0, -1, 0), (1, -1, 6),
        (-1, 0, -6), (1, 0, 6),
        (-1, 1, -6), (0, 1, 0), (1, 1, 6),
        (0, 0, 0)
    ]

    for dw, dt, dz in deltas[:sample_size]:
        if dw == dt == dz == 0:
            continue

        neighbor_weather = int(np.clip(state[0] + dw, MIN_X, MAX_X))
        neighbor_time = int(np.clip(state[1] + dt, MIN_Y, MAX_Y))
        neighbor_z = int(np.clip(state[2] + dz, MIN_Z, MAX_Z))
        neighbor = (neighbor_weather, neighbor_time, neighbor_z)

        n_trig = section2_flow_pressure_density_hybrid(neighbor[0], neighbor[1], neighbor[2])
        if not n_trig:
            continue

        rob += jaccard_similarity(n_trig, base)
        neighbors += 1

    return rob / neighbors if neighbors > 0 else 0.0

def generate_samples_for_all_paths(num_candidates=2000, top_k=200, run_id=1):
    """Path """
    BEST_WEIGHTS = [0.55, 0.25, 0.2]

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_individual.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Individual Path {path_id} (Weighted Screening) - Run {run_id}\n")
            f.write("weather time_period z\tScore\tSimilarity\tRobustness\tLengthDiff\n")
            for s in samples:
                weather, time_period, z = s['state']
                f.write(
                    f"{weather} {time_period} {z}\t{s['score']:.4f}\t{s['similarity']:.4f}\t"
                    f"{s['robustness']:.4f}\t{s['length_diff']:.4f}\n"
                )

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_individual"

    for path_idx in range(len(target_paths)):
        path = target_paths[path_idx]
        candidate_samples = []
        attempts = 0
        max_attempts = num_candidates * 10

        while len(candidate_samples) < num_candidates and attempts < max_attempts:
            attempts += 1

            weather = np.random.randint(MIN_X, MAX_X + 1)
            time_period = np.random.randint(MIN_Y, MAX_Y + 1)
            z = np.random.randint(MIN_Z, MAX_Z + 1)
            state = (weather, time_period, z)

            triggered = section2_flow_pressure_density_hybrid(weather, time_period, z)
            if not triggered:
                continue

            sim = jaccard_similarity(triggered, path)
            rob = compute_robustness(state, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))

            candidate_samples.append({
                'state': state,
                'similarity': sim,
                'robustness': rob,
                'length_diff': len_diff,
                'triggered': triggered
            })

        if candidate_samples:
            for sample in candidate_samples:
                score = (BEST_WEIGHTS[0] * sample['similarity'] +
                         BEST_WEIGHTS[1] * sample['robustness'] +
                         BEST_WEIGHTS[2] * sample['length_diff'])
                sample['score'] = score

            candidate_samples.sort(key=lambda x: x['score'], reverse=True)
            selected_samples = candidate_samples[:top_k]
            save_samples(path_id=path_idx + 1, samples=selected_samples, base_dir=base_dir)

# === Shared Experience Replay ===
class SharedExperienceReplay:
    def __init__(self, capacity=20000):
        self.capacity = capacity
        self.buffer = deque(maxlen=self.capacity)
        self.priorities = deque(maxlen=self.capacity)

    def append(self, experience):
        self.buffer.append(experience)
        self.priorities.append(experience[-1])

    def sample(self, batch_size, alpha=0.6):
        if len(self.buffer) < batch_size:
            return [], [], []

        priorities = np.array(self.priorities, dtype=np.float64)
        priorities = np.power(priorities, alpha)
        probabilities = priorities / np.sum(priorities)

        batch_indices = np.random.choice(len(self.buffer), batch_size, p=probabilities, replace=False)
        batch = [self.buffer[idx] for idx in batch_indices]

        return batch, batch_indices, probabilities[batch_indices]

    def update_priorities(self, batch_indices, td_errors):
        for idx, td_error in zip(batch_indices, td_errors):
            if idx < len(self.priorities):
                self.priorities[idx] = max(abs(td_error), 1e-6)

    def __len__(self):
        return len(self.buffer)

    def get_high_reward_samples(self, target_path, num_samples=20):
        if len(self.buffer) == 0:
            return []

        samples_with_scores = []
        seen_states = set()

        for experience in self.buffer:
            state_tensor = experience[0]
            state_norm = state_tensor.cpu().numpy().flatten()
            state_tuple = denormalize_state((state_norm[0], state_norm[1], state_norm[2]))

            if state_tuple in seen_states:
                continue
            seen_states.add(state_tuple)

            triggered = section2_flow_pressure_density_hybrid(state_tuple[0], state_tuple[1], state_tuple[2])
            reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)

            samples_with_scores.append((state_tuple, reward, sim, triggered))

        samples_with_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_scores[:num_samples]

def load_path_data(file_path):
    path_data = []

    if not os.path.exists(file_path):
        print(f":  {file_path}")
        return path_data

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
            for line in lines[2:]:
                parts = line.strip().split("\t")
                if parts:
                    values = parts[0].split()
                    if len(values) >= 3:
                        state = (int(values[0]), int(values[1]), int(values[2]))
                        path_data.append(state)
    except Exception as e:
        print(f" {file_path}: {e}")

    return path_data

# === DQN ===
class DQN(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dims=[128, 64]):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, hidden_dims[0])
        self.fc2 = nn.Linear(hidden_dims[0], hidden_dims[1])
        self.fc3 = nn.Linear(hidden_dims[1], action_dim)
        self.dropout = nn.Dropout(0.1)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = self.dropout(x)
        x = torch.relu(self.fc2(x))
        return self.fc3(x)

# === DQN Agent with PER ===
class DQNAgentWithPER:
    def __init__(self, state_dim, action_dim, replay_buffer,
                 gamma=0.99, epsilon=1.0, epsilon_decay=0.995,
                 epsilon_min=0.1, learning_rate=0.001, alpha=0.6, beta=0.4):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer
        self.alpha = alpha
        self.beta = beta

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        """
        ()
        30, :
        - weather: +/-1, 0(x2)
        - time_period: +/-1, 0(x2)
        - z: +/-12(20%), +/-6(10%), +/-3(5%), 0(x2)
        """
        delta_values_weather_time = [1, 0, 0, -1]
        delta_values_z = [12, 6, 3, 0, 0, -3, -6, -12]

        dim = action_idx // 10
        delta_idx = action_idx % 10

        if dim == 0:  # weather
            if delta_idx >= 4:
                delta_idx = 3
            return (delta_values_weather_time[delta_idx], 0, 0)
        elif dim == 1:  # time_period
            if delta_idx >= 4:
                delta_idx = 3
            return (0, delta_values_weather_time[delta_idx], 0)
        elif dim == 2:  # z
            if delta_idx >= 8:
                delta_idx = 7
            return (0, 0, delta_values_z[delta_idx])

    def act(self, state_norm, legal_actions=None):
        if legal_actions is None:
            legal_actions = list(range(self.action_dim))

        if not legal_actions:
            return None

        if random.random() < self.epsilon:
            return random.choice(legal_actions)

        state_tensor = torch.tensor(state_norm, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)[0]

        legal_q_values = q_values[legal_actions]
        best_legal_idx = torch.argmax(legal_q_values).item()
        return legal_actions[best_legal_idx]

    def get_legal_actions(self, state):
        legal_actions = []

        for action_idx in range(self.action_dim):
            dw, dt, dz = self.decode_action(action_idx)

            next_weather = state[0] + dw
            next_time = state[1] + dt
            next_z = state[2] + dz

            if (MIN_X <= next_weather <= MAX_X and
                    MIN_Y <= next_time <= MAX_Y and
                    MIN_Z <= next_z <= MAX_Z):
                legal_actions.append(action_idx)

        return legal_actions

    def store_transition(self, state, action, reward, next_state, done):
        state_norm = normalize_state(state)
        next_state_norm = normalize_state(next_state)

        state_tensor = torch.tensor(state_norm, dtype=torch.float32).unsqueeze(0).to(device)
        next_state_tensor = torch.tensor(next_state_norm, dtype=torch.float32).unsqueeze(0).to(device)

        with torch.no_grad():
            q_values = self.model(state_tensor)
            next_q_values = self.target_model(next_state_tensor)
            max_next_q = next_q_values.max(1)[0]
            target_q = reward + (self.gamma * max_next_q * (1 - done))
            td_error = abs(q_values[0][action].item() - target_q.item())

        self.replay_buffer.append((state_tensor, action, reward, next_state_tensor, done, td_error))
        return td_error

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return 0.0

        batch, batch_indices, probabilities = self.replay_buffer.sample(batch_size, alpha=self.alpha)

        if not batch:
            return 0.0

        states, actions, rewards, next_states, dones, _ = zip(*batch)

        weights = (len(self.replay_buffer) * probabilities) ** (-self.beta)
        weights = weights / weights.max()
        weights = torch.tensor(weights, dtype=torch.float32).to(device)

        states = torch.cat(states).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.cat(next_states).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q = self.target_model(next_states).max(1)[0].detach()
        target_q = rewards + (self.gamma * next_max_q * (1 - dones))

        td_errors = current_q - target_q
        weighted_loss = (td_errors.pow(2) * weights).mean()

        self.optimizer.zero_grad()
        weighted_loss.backward()
        torch.nn.utils.clip_grad_norm_(self.model.parameters(), 1.0)
        self.optimizer.step()

        new_priorities = abs(td_errors.detach().cpu().numpy())
        self.replay_buffer.update_priorities(batch_indices, new_priorities)

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

        return weighted_loss.item()

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())

# === 训练函数 ===
def generate_and_train_for_individual_paths(path_documents, repeats=5, batch_size=32, run_id=1):
    state_dim = 3
    action_dim = 30

    shared_replay_buffer = SharedExperienceReplay(capacity=20000)
    agent = DQNAgentWithPER(state_dim, action_dim, shared_replay_buffer)

    total_cumulative_reward = 0
    path_rewards = {}

    print(f"\n===  {run_id}/20 Start training(+)===")
    start_time = time.time()

    SAMPLES_PER_BATCH = 50
    NUM_BATCHES = 4
    STEPS_PER_SAMPLE = 3

    for path_idx in range(len(target_paths)):
        path_id = path_idx + 1
        print(f"\n{'=' * 60}")
        print(f"Path  {path_id}/{NUM_PATHS}")
        print(f"{'=' * 60}")

        file_path = os.path.join(path_documents, f"path{path_id}_individual.txt")
        if not os.path.exists(file_path):
            print(f"  :  {file_path}")
            continue

        path_data = load_path_data(file_path)
        if not path_data:
            print(f"  : Path  {path_id} ")
            continue

        target_path = target_paths[path_idx]

        if path_idx not in path_rewards:
            path_rewards[path_idx] = 0

        for repeat_idx in range(repeats):
            print(f"\n  Run  {repeat_idx + 1}/{repeats} ")

            for batch_idx in range(NUM_BATCHES):
                batch_start = batch_idx * SAMPLES_PER_BATCH
                batch_end = min(batch_start + SAMPLES_PER_BATCH, len(path_data))

                print(f"     {batch_idx + 1}/{NUM_BATCHES} ( {batch_start}-{batch_end})")

                for sample_idx in range(batch_start, batch_end):
                    state = path_data[sample_idx]
                    prev_state = None
                    prev_triggered = None

                    for step in range(STEPS_PER_SAMPLE):
                        legal_actions = agent.get_legal_actions(state)

                        if not legal_actions:
                            break

                        state_norm = normalize_state(state)
                        action = agent.act(state_norm, legal_actions)
                        if action is None:
                            break

                        dw, dt, dz = agent.decode_action(action)

                        next_state = (
                            int(np.clip(state[0] + dw, MIN_X, MAX_X)),
                            int(np.clip(state[1] + dt, MIN_Y, MAX_Y)),
                            int(np.clip(state[2] + dz, MIN_Z, MAX_Z))
                        )

                        triggered = section2_flow_pressure_density_hybrid(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered,
                                                prev_triggered, prev_state)
                        done = (step == STEPS_PER_SAMPLE - 1)

                        agent.store_transition(state, action, reward, next_state, done)

                        prev_state = state
                        prev_triggered = triggered
                        state = next_state

                        total_cumulative_reward += reward
                        path_rewards[path_idx] += reward

                if len(agent.replay_buffer) >= batch_size:
                    loss = agent.train(batch_size)
                    print(f"       {batch_idx + 1} completed, : {loss:.4f}")

                if (batch_idx + 1) % 2 == 0:
                    agent.update_target_model()
                    print(f"      completed {batch_idx + 1} , ")

        print(f"\nPath  {path_id} completed, : {path_rewards[path_idx]:.2f}")
        print(f"Shared Buffer Size: {len(shared_replay_buffer)}")

    training_time = time.time() - start_time
    print(f"\n===  {run_id}/20 completed, : {training_time:.2f} seconds ===")

    return agent, shared_replay_buffer, total_cumulative_reward, path_rewards, training_time

# === Excel 报告生成 ===
def create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir):
    os.makedirs(output_dir, exist_ok=True)

    similar_group_paths = [idx + 1 for idx in similar_group]
    isolated_group_paths = [idx + 1 for idx in isolated_group]

    wb = Workbook()

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    header_color = "4472C4"
    similar_group_color = "E2EFDA"
    isolated_group_color = "FCE4D6"
    stats_color = "FFF2CC"

    # === 1: Path  ===
    ws_paths = wb.active
    ws_paths.title = "Path "

    path_headers = ['Path ID', ''] + [f'Run {i}' for i in range(1, 21)] + \
                   ['Average Similarity', 'Maximum Similarity', 'Minimum Similarity', 'Standard deviation']

    for col, header in enumerate(path_headers, 1):
        cell = ws_paths.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_paths.row_dimensions[1].height = 30

    for path_id in range(1, NUM_PATHS + 1):
        row = path_id + 1

        if path_id in similar_group_paths:
            group_type = "High-correlation path group"
            row_color = similar_group_color
        elif path_id in isolated_group_paths:
            group_type = "Low-correlation path group"
            row_color = isolated_group_color
        else:
            group_type = "Ungrouped"
            row_color = "FFFFFF"

        cell = ws_paths.cell(row=row, column=1, value=f"Path {path_id}")
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        cell.border = thin_border

        cell = ws_paths.cell(row=row, column=2, value=group_type)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        cell.border = thin_border

        path_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            sim = run_data['path_similarities'].get(path_id, {}).get('avg_similarity', 0.0)
            path_similarities.append(sim)

            cell = ws_paths.cell(row=row, column=3 + run_idx, value=round(sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        stats_values = [
            round(np.mean(path_similarities), 4),
            round(np.max(path_similarities), 4),
            round(np.min(path_similarities), 4),
            round(np.std(path_similarities), 4)
        ]

        for i, value in enumerate(stats_values):
            cell = ws_paths.cell(row=row, column=23 + i, value=value)
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
            cell.font = Font(bold=True, size=10)
            cell.border = thin_border

    ws_paths.column_dimensions['A'].width = 13
    ws_paths.column_dimensions['B'].width = 16
    for col in range(3, 23):
        ws_paths.column_dimensions[get_column_letter(col)].width = 10
    for col in range(23, 27):
        ws_paths.column_dimensions[get_column_letter(col)].width = 13

    # === 2:  ===
    ws_groups = wb.create_sheet("")

    group_headers = ['Group Name', 'Included Paths'] + [f'Run {i}' for i in range(1, 21)] + \
                    ['Average Similarity', 'Standard deviation']

    for col, header in enumerate(group_headers, 1):
        cell = ws_groups.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_groups.row_dimensions[1].height = 30

    row = 2

    # High-correlation path group
    cell = ws_groups.cell(row=row, column=1, value="High-correlation path group")
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    cell.border = thin_border

    cell = ws_groups.cell(row=row, column=2, value=','.join(map(str, similar_group_paths)))
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    cell.border = thin_border

    group_similarities = []
    for run_idx, run_data in enumerate(all_runs_data):
        group_sim = np.mean([
            run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
            for p in similar_group_paths
        ])
        group_similarities.append(group_sim)

        cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(group_sim, 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    cell = ws_groups.cell(row=row, column=23, value=round(np.mean(group_similarities), 4))
    cell.number_format = '0.0000'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border

    cell = ws_groups.cell(row=row, column=24, value=round(np.std(group_similarities), 4))
    cell.number_format = '0.0000'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border

    row += 1

    # Low-correlation path group
    if isolated_group_paths:
        cell = ws_groups.cell(row=row, column=1, value="Low-correlation path group")
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        cell.border = thin_border

        cell = ws_groups.cell(row=row, column=2, value=','.join(map(str, isolated_group_paths)))
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        cell.border = thin_border

        isolated_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            iso_sim = np.mean([
                run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
                for p in isolated_group_paths
            ])
            isolated_similarities.append(iso_sim)

            cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(iso_sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        cell = ws_groups.cell(row=row, column=23, value=round(np.mean(isolated_similarities), 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border

        cell = ws_groups.cell(row=row, column=24, value=round(np.std(isolated_similarities), 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border

    ws_groups.column_dimensions['A'].width = 16
    ws_groups.column_dimensions['B'].width = 22
    for col in range(3, 23):
        ws_groups.column_dimensions[get_column_letter(col)].width = 10
    ws_groups.column_dimensions[get_column_letter(23)].width = 14
    ws_groups.column_dimensions[get_column_letter(24)].width = 12

    # === 3: Detailed Sample Data ===
    ws_samples = wb.create_sheet("Detailed Sample Data")

    sample_headers = ['Run', 'Path ID', 'Sample ID', '', '',
                      '', 'Similarity', 'Triggered Rule Set']

    for col, header in enumerate(sample_headers, 1):
        cell = ws_samples.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_samples.row_dimensions[1].height = 30

    sample_row = 2
    for run_idx, run_data in enumerate(all_runs_data, 1):
        for path_id in range(1, NUM_PATHS + 1):
            samples = run_data['path_samples'].get(path_id, [])

            if path_id in similar_group_paths:
                path_color = similar_group_color
            elif path_id in isolated_group_paths:
                path_color = isolated_group_color
            else:
                path_color = "FFFFFF"

            for sample_idx, (state_tuple, reward, sim, triggered) in enumerate(samples, 1):
                weather, time_period, z = state_tuple
                triggered_str = ','.join(map(str, sorted(triggered)))

                cell = ws_samples.cell(row=sample_row, column=1, value=f"Run {run_idx}")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=path_color, end_color=path_color, fill_type="solid")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=2, value=f"Path {path_id}")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=path_color, end_color=path_color, fill_type="solid")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=3, value=sample_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                for col_offset, value in enumerate([weather, time_period, z]):
                    cell = ws_samples.cell(row=sample_row, column=4 + col_offset, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=7, value=round(sim, 4))
                cell.number_format = '0.0000'
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=8, value=f"{{{triggered_str}}}")
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border

                sample_row += 1

    sample_widths = [13, 13, 11, 10, 10, 12, 12, 45]
    for i, width in enumerate(sample_widths, 1):
        ws_samples.column_dimensions[get_column_letter(i)].width = width

    # === 4: Run Statistics Summary ===
    ws_summary = wb.create_sheet("Run Statistics Summary")

    summary_headers = ['Run', 'Training Time( seconds)', 'Overall Average Similarity', 'Maximum Similarity',
                       'Minimum Similarity', '', '', 'Shared Buffer Size']

    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_summary.row_dimensions[1].height = 30

    for run_idx, run_data in enumerate(all_runs_data, 1):
        row = run_idx + 1

        # Average Similarity
        high_group_avg = np.mean([
            run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
            for p in similar_group_paths
        ])

        low_group_avg = 0.0
        if isolated_group_paths:
            low_group_avg = np.mean([
                run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
                for p in isolated_group_paths
            ])

        values = [
            f"Run {run_idx}",
            round(run_data['training_time'], 2),
            round(run_data['overall_avg_similarity'], 4),
            round(run_data['max_similarity'], 4),
            round(run_data['min_similarity'], 4),
            round(high_group_avg, 4),
            round(low_group_avg, 4),
            20000
        ]

        for col, value in enumerate(values, 1):
            cell = ws_summary.cell(row=row, column=col, value=value)
            if col == 1:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 2:
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col >= 3 and col <= 7:
                cell.number_format = '0.0000'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    #
    stat_row = len(all_runs_data) + 2
    stat_labels = ['', '/', '', '', '', '', '', '']

    for col, label in enumerate(stat_labels, 1):
        cell = ws_summary.cell(row=stat_row, column=col, value=label)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    stat_row += 1

    #
    training_times = [r['training_time'] for r in all_runs_data]
    overall_avgs = [r['overall_avg_similarity'] for r in all_runs_data]
    max_sims = [r['max_similarity'] for r in all_runs_data]
    min_sims = [r['min_similarity'] for r in all_runs_data]

    high_group_avgs = []
    low_group_avgs = []
    for run_data in all_runs_data:
        high_avg = np.mean([
            run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
            for p in similar_group_paths
        ])
        high_group_avgs.append(high_avg)

        if isolated_group_paths:
            low_avg = np.mean([
                run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
                for p in isolated_group_paths
            ])
            low_group_avgs.append(low_avg)

    stat_values = [
        '',
        round(np.sum(training_times), 2),
        round(np.mean(overall_avgs), 4),
        round(np.max(max_sims), 4),
        round(np.min(min_sims), 4),
        round(np.mean(high_group_avgs), 4),
        round(np.mean(low_group_avgs), 4) if low_group_avgs else 0.0,
        20000
    ]

    for col, value in enumerate(stat_values, 1):
        cell = ws_summary.cell(row=stat_row, column=col, value=value)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        if col == 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col == 2:
            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col >= 3 and col <= 7:
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    summary_widths = [13, 16, 18, 14, 14, 16, 16, 14]
    for i, width in enumerate(summary_widths, 1):
        ws_summary.column_dimensions[get_column_letter(i)].width = width

    output_path = os.path.join(output_dir, "20 run_.xlsx")
    wb.save(output_path)
    print(f"\n Consolidated Excel report generated: {output_path}")
    print(f"   4: Path , , Detailed Sample Data, Run Statistics Summary")

def run_20_times_training():
    """20"""
    model_path_base = r"D:\Experiment\CNN\DQNNEW\saved_models_new_vars"
    path_documents = r"D:\Experiment\CNN\DQNNEW\path_samples_individual"
    output_dir = r"D:\Experiment\CNN\ComparisonExperiment2\excel_reports_new_vars"

    os.makedirs(model_path_base, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    similar_group, isolated_group = group_paths_by_similarity(target_paths)

    similar_group_display = [idx + 1 for idx in similar_group]
    isolated_group_display = [idx + 1 for idx in isolated_group]

    print("=" * 60)
    print("20 - ")
    print("=" * 60)
    print(f"\nPath : {NUM_PATHS} ")
    print(f"\nAutomatic grouping results:")
    print(f"  Similar path group: {similar_group_display}")
    print(f"  Isolated path group: {isolated_group_display}")
    print(f"\n:")
    print(f"   (weather): {MIN_X}-{MAX_X}")
    print(f"   (time_period): {MIN_Y}-{MAX_Y}")
    print(f"   (z): {MIN_Z}-{MAX_Z}")
    print(f"\n:")
    print(f"  x = z ()")
    print(f"  y = (weather * time_period * 10 + z) % 100 + 1")
    print(f"\n:")
    print(f"  : [0, 1]")
    print(f"  : ")
    print(f"  : (seen_states)")
    print(f"\n (30):")
    print(f"  weather: +/-1, 0(x2)")
    print(f"  time_period: +/-1, 0(x2)")
    print(f"  z: +/-12, +/-6, +/-3, 0(x2)")
    print("\n" + "=" * 60)

    all_runs_data = []
    total_start_time = time.time()

    for run_id in range(1, 21):
        print(f"\n{'=' * 60}")
        print(f"Start run  {run_id}/20  run")
        print(f"{'=' * 60}")

        print(f"[Run {run_id}] Generating samples...")
        generate_samples_for_all_paths(num_candidates=2000, top_k=200, run_id=run_id)

        print(f"[Run {run_id}] Start training...")
        agent, shared_buffer, total_cumulative_reward, path_rewards, training_time = \
            generate_and_train_for_individual_paths(path_documents, repeats=5,
                                                    batch_size=32, run_id=run_id)

        model_path = os.path.join(model_path_base, f"trained_model_run_{run_id}.pth")
        torch.save({
            'model_state_dict': agent.model.state_dict(),
            'optimizer_state_dict': agent.optimizer.state_dict(),
            'epsilon': agent.epsilon,
            'run_id': run_id,
            'normalized': True,
            'value_ranges': {
                'weather': [MIN_X, MAX_X],
                'time_period': [MIN_Y, MAX_Y],
                'z': [MIN_Z, MAX_Z]
            }
        }, model_path)
        print(f"[Run {run_id}] Model saved: {model_path}")

        run_data = {
            'run_id': run_id,
            'training_time': training_time,
            'total_reward': total_cumulative_reward,
            'path_rewards': path_rewards,
            'path_similarities': {},
            'path_samples': {}
        }

        all_similarities = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            high_reward_samples = shared_buffer.get_high_reward_samples(target_path, num_samples=20)

            if high_reward_samples:
                similarities = [sim for _, _, sim, _ in high_reward_samples]
                run_data['path_similarities'][path_idx + 1] = {
                    'avg_similarity': np.mean(similarities),
                    'max_similarity': np.max(similarities),
                    'min_similarity': np.min(similarities),
                    'sample_count': len(similarities)
                }
                run_data['path_samples'][path_idx + 1] = high_reward_samples
                all_similarities.extend(similarities)
            else:
                run_data['path_similarities'][path_idx + 1] = {
                    'avg_similarity': 0.0,
                    'max_similarity': 0.0,
                    'min_similarity': 0.0,
                    'sample_count': 0
                }
                run_data['path_samples'][path_idx + 1] = []

        if all_similarities:
            run_data['overall_avg_similarity'] = np.mean(all_similarities)
            run_data['max_similarity'] = np.max(all_similarities)
            run_data['min_similarity'] = np.min(all_similarities)
        else:
            run_data['overall_avg_similarity'] = 0.0
            run_data['max_similarity'] = 0.0
            run_data['min_similarity'] = 0.0

        all_runs_data.append(run_data)

        print(f"[Run {run_id}] completed! Overall Average Similarity: {run_data['overall_avg_similarity']:.4f}")
        print(f"{'=' * 60}\n")

    total_time = time.time() - total_start_time

    print("\nGenerating consolidated Excel report...")
    create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir)

    print("\n" + "=" * 60)
    print(f"\nAll results have been saved to: {output_dir}")
    print("=" * 60)

if __name__ == "__main__":
    run_20_times_training()