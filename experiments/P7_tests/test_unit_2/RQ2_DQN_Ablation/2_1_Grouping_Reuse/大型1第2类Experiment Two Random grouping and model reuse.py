import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import math

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === 修改取值范围为 X:1~300, Y:1~200, Z:1~5 ===
STATE_RANGES = {
    'dx': (1, 300),
    'dy': (1, 200),
    'dz': (1, 5),
}
STATE_NAMES = ('dx', 'dy', 'dz')
STATE_MIN = np.array([STATE_RANGES[name][0] for name in STATE_NAMES], dtype=np.int32)
STATE_MAX = np.array([STATE_RANGES[name][1] for name in STATE_NAMES], dtype=np.int32)


def clip_state(state):
    """ dx/dy/dz ,  int tuple."""
    return tuple(np.clip(np.array(state, dtype=np.int32), STATE_MIN, STATE_MAX).astype(int))


def random_state():
    """ dx/dy/dz ."""
    return tuple(random.randint(STATE_RANGES[name][0], STATE_RANGES[name][1]) for name in STATE_NAMES)


class StateNormalizer:
    """:  dx/dy/dz  [0, 1]."""

    def __init__(self, ranges=None):
        self.ranges = ranges or STATE_RANGES
        self.names = STATE_NAMES

    def normalize(self, state):
        """ -> ."""
        state = np.array(state, dtype=np.float32)
        normalized = np.zeros_like(state, dtype=np.float32)
        for i, name in enumerate(self.names):
            low, high = self.ranges[name]
            normalized[i] = (state[i] - low) / (high - low)
        return normalized

    def denormalize(self, normalized_state):
        """ ->  dx/dy/dz, ."""
        normalized_state = np.array(normalized_state, dtype=np.float32)
        denormalized = np.zeros_like(normalized_state, dtype=np.float32)
        for i, name in enumerate(self.names):
            low, high = self.ranges[name]
            denormalized[i] = normalized_state[i] * (high - low) + low
            denormalized[i] = np.clip(np.round(denormalized[i]), low, high).astype(int)
        return denormalized


normalizer = StateNormalizer()


# === reward function ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward


# ========== 规则触发函数（section2_flow_pressure_density_hybrid） ==========
def section2_flow_pressure_density_hybrid(x, y, z):
    b = {}  # 使用字典存储触发的规则

    if (30 < x < 170) != (30 < x < 100): b[0] = 1
    if (30 < x < 170) != (30 < x < 190): b[1] = 2
    if (30 < x < 170) != (30 < x < 140): b[2] = 3
    if (30 < x < 170) != (30 < x < 150): b[3] = 4
    if (30 < x < 170) != (30 < x < 160): b[4] = 5
    if (100 < y < 200) != (150 < y < 200): b[5] = 6
    if (100 < y < 200) != (50 < y < 200): b[6] = 7
    if (100 < y < 200) != (100 < y < 150): b[7] = 8
    if (1 < z < 4) != (1 < z < 3): b[8] = 9
    if (x > 10 and x < 152) != (x > 10 and x < 100): b[9] = 10

    # 规则11-20
    if (x > 40 and x < 152) != (x > 40 and x < 202): b[10] = 11
    if (abs(x - 50) < 4) != (abs(x - 50) < 2): b[11] = 12
    if (abs(x - 50) < 4) != (abs(x + 50) < 4): b[12] = 13
    if (abs(x - 50) < 4) != (abs(x * 50) < 4): b[13] = 14
    if (x > 135) != (x > 215): b[14] = 15
    if (x > 135) != (x > 235): b[15] = 16
    if (x > 135) != (x > 225): b[16] = 17
    if (y > 100) != (y > 120): b[17] = 18
    if (y > 100) != (y > 140): b[18] = 19
    if (y > 100) != (y > 50): b[19] = 20

    # 规则21-30
    if (x < 165) != (x < 265): b[20] = 21
    if (x < 165) != (x < 245): b[21] = 22
    if (x < 165) != (x < 262): b[22] = 23
    if (y < 150) != (y < 130): b[23] = 24
    if (y < 150) != (x < 100): b[24] = 25
    if (y < 190) != (y < 120): b[25] = 26
    if ((x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 125) != (
            (x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 115): b[26] = 27
    if ((x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 125) != (
            (x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 85): b[27] = 28
    if ((x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 125) != (
            (x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 25): b[28] = 29
    if (x / (y / 30) > 1 and x / (y / 30) < 2) != (x / (y / 30) > 1 and x / (y / 30) < 4): b[29] = 30

    # 规则31-40
    if (x / (y / 30) > 1 and x / (y / 30) < 2) != (x / (y / 30) > 1 and x / (y / 30) < 3): b[30] = 31
    if (x / (z / 2) > 2 and x / (z / 2) < 4) != (x / (z / 2) > 1 and x / (z / 2) < 4): b[31] = 32
    if (x / (z / 2) > 2 and x / (z / 2) < 4) != (x / (z / 3) > 2 and x / (z / 2) < 4): b[32] = 33
    if (x / (z / 2) > 2 and x / (z / 2) < 4) != (x / (z / 2) > 2 and x / (z / 3) < 4): b[33] = 34
    if ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 55) != ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 155): b[
        34] = 35
    if ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 55) != ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 125): b[
        35] = 36
    if ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 55) != ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 25): b[
        36] = 37
    if (x - y / 30 > -10 and x - y / 30 < 10) != (x - y / 30 > -10 and x - y / 30 < 5): b[37] = 38
    if (x - y / 30 > -10 and x - y / 30 < 10) != (x - y / 30 > -10 and x - y / 30 < 6): b[38] = 39
    if (x - y / 30 > -10 and x - y / 30 < 10) != (x - y / 30 > -10 and x - y / 30 < 8): b[39] = 40

    # 规则41-50
    if (abs(x / (y / 30) - 1.0) < 1) != (abs(x / (y / 30) + 1.0) < 1): b[40] = 41
    if (abs(x / (y / 30) - 1.0) < 1) != (abs(x / (y / 30) - 1.0) < 2): b[41] = 42
    if (abs(x / (y / 30) - 1.0) < 1) != (abs(x / (y / 30) - 1.0) < 3): b[42] = 43
    if (x / ((y / 30) + 20) > 1 and x / ((y / 30) + 20) < 2) != (x / ((y / 30) + 20) > 1 and x / ((y / 20) + 20) < 2):
        b[43] = 44
    if (x / ((y / 30) + 20) > 1 and x / ((y / 30) + 20) < 2) != (x / ((y / 30) + 20) > 1 and x / ((y / 40) + 20) < 2):
        b[44] = 45
    if (x / ((y / 30) + 20) > 1 and x / ((y / 30) + 20) < 2) != (x / ((y / 30) + 20) > 1 and x / ((y / 30) + 40) < 2):
        b[45] = 46
    if ((z / 50) / (x / 50) > 1 and (z / 50) / (x / 50) < 3) != ((z / 50) / (x / 50) > 1 and (z / 50) / (x / 50) < 4):
        b[46] = 47
    if ((z / 50) / (x / 50) > 1 and (z / 50) / (x / 50) < 3) != ((z / 50) / (x / 50) > 2 and (z / 50) / (x / 50) < 3):
        b[47] = 48
    if ((z / 50) / (x / 50) > 1 and (z / 50) / (x / 50) < 3) != ((z / 50) / (x / 50) > 3 and (z / 50) / (x / 50) < 3):
        b[48] = 49
    if ((x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 45) != (
            (x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 70): b[49] = 50

    # 规则51-60
    if ((x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 45) != (
            (x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 60): b[50] = 51
    if ((x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 45) != (
            (x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 6 < 45): b[51] = 52
    if (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 8) != (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 48): b[
        52] = 53
    if (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 8) != (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 50): b[
        53] = 54
    if (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 8) != (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 58): b[
        54] = 55
    if ((x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 200) != (
            (x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 400): b[55] = 56
    if ((x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 200) != (
            (x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 300): b[56] = 57
    if ((x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 200) != (
            (x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 500): b[57] = 58
    if (abs((x + y / 30) - 100) < 15) != (abs((x + y / 30) - 100) < 35): b[58] = 59
    if (abs((x + y / 30) - 100) < 15) != (abs((x + y / 30) - 100) < 45): b[59] = 60

    # 规则61-70
    if (abs((x + y / 30) - 100) < 15) != (abs((x + y / 30) - 100) < 10): b[60] = 61
    if (x / 50 > 1 and x / 50 < 4) != (x / 50 > 3 and x / 50 < 4): b[61] = 62
    if (x / 50 > 1 and x / 50 < 4) != (x / 50 > 1 and x / 50 < 5): b[62] = 63
    if (x < 135) != (x < 215): b[63] = 64
    if (x < 135) != (x < 225): b[64] = 65
    if (x < 135) != (x < 225): b[65] = 66
    if (x > 65) != (x > 165): b[66] = 67
    if (x > 65) != (x > 215): b[67] = 68
    if (x > 65) != (x > 251): b[68] = 69
    if (y < 110) != (y < 80): b[69] = 70

    # 规则71-80
    if (y < 110) != (y < 50): b[70] = 71
    if (y < 110) != (y < 119): b[71] = 72
    if (x < 32 or x > 68) != (x < 32 or x > 168): b[72] = 73
    if (x < 32 or x > 68) != (x < 32 or x > 118): b[73] = 74
    if (x < 32 or x > 68) != (x < 32 or x > 218): b[74] = 75
    if (x < 35 and y < 120) != (x < 135 and y < 120): b[75] = 76
    if (x < 35 and y < 120) != (x < 235 and y < 120): b[76] = 77
    if (x < 35 and y < 120) != (x < 35 and y < 60): b[77] = 78
    if (abs(x / (y / 30) - 1.0) > 1) != (abs(x / (y / 30) - 3.0) > 1): b[78] = 79
    if (abs(x / (y / 30) - 1.0) > 1) != (abs(x / (y / 30) - 1.0) > 2): b[79] = 80

    # 规则81-90
    if (abs(x / (y / 30) - 1.0) > 1) != (abs(x / (y / 30) - 1.0) > 3): b[80] = 81
    if ((x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 135) != (
            (x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 235): b[81] = 82
    if ((x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 135) != (
            (x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 335): b[82] = 83
    if ((x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 135) != (
            (x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 435): b[83] = 84
    if (x < 30 or x > 70) != (x < 30 or x * 7 > 70): b[84] = 85
    if (x < 30 or x > 70) != (x < 30 or x > 270): b[85] = 86
    if (x < 30 or x > 70) != (x < 30 or x > 170): b[86] = 87
    if (x / (y / 30) > 1 and x / (y / 30) < 3) != (x / (y / 30) > 1 and x / (y / 20) < 3): b[87] = 88
    if (x / (y / 30) > 1 and x / (y / 30) < 3) != (x / (y / 30) > 1 and x / (y / 30) < 4): b[88] = 89
    if (x / (y / 30) > 1 and x / (y / 30) < 3) != (x / (y / 20) > 1 and x / (y / 30) < 3): b[89] = 90

    # 规则91-100
    if ((x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 112) != (
            (x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 62): b[90] = 91
    if ((x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 112) != (
            (x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 92): b[91] = 92
    if ((x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 112) != (
            (x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 82): b[92] = 93
    if (abs((x + y / 30) / 2 - 50) < 1) != (abs((x + y / 30) / 2 - 50) < 4): b[93] = 94
    if (abs((x + y / 30) / 2 - 50) < 1) != (abs((x + y / 30) / 2 + 50) < 1): b[94] = 95
    if (abs((x + y / 30) / 2 - 50) < 1) != (abs((x + y / 20) / 2 + 50) < 1): b[95] = 96
    if (x / (z / 2) > 2 and x / (z / 2) < 4) != (x / (z / 2) > 1 and x / (z / 2) < 4): b[96] = 97
    if (x / (z / 2) > 20 and x / (z / 2) < 100) != (x / (z / 3) > 20 and x / (z / 2) < 100): b[97] = 98
    if (x / (z / 2) > 20 and x / (z / 2) < 100) != (x / (z / 2) > 20 and x / (z / 2) < 150): b[98] = 99
    if ((y / 30) / (z / 50) > 38 and (y / 30) / (z / 50) < 42) != (
            (y / 30) / (z / 50) > 48 and (y / 30) / (z / 50) < 42): b[99] = 100

    # 规则101-110
    if ((y / 30) / (z / 50) > 38 and (y / 30) / (z / 50) < 42) != (
            (y / 30) / (z / 50) > 58 and (y / 30) / (z / 50) < 42): b[100] = 101
    if ((y / 30) / (z / 50) > 38 and (y / 30) / (z / 50) < 42) != (
            (y / 30) / (z / 50) > 38 and (y / 30) / (z / 50) < 72): b[101] = 102
    if ((x - 40) / 20 > 1 and (x - 40) / 20 < 6) != ((x - 40) / 30 > 1 and (x - 40) / 20 < 7): b[102] = 103
    if ((x - 40) / 20 > 1 and (x - 40) / 20 < 6) != ((x - 50) / 20 > 1 and (x - 40) / 20 < 7): b[103] = 104
    if ((x - 40) / 20 > 1 and (x - 40) / 20 < 6) != ((x - 40) / 20 > 1 and (x - 50) / 20 < 7): b[104] = 105

    triggered = set(b.values())
    return triggered


# === 将规则函数赋给 execute_Tr，供后续使用（修正为 section2） ===
execute_Tr = section2_flow_pressure_density_hybrid

# === 目标路径组（匹配 section2 的规则编号 1~105） ===
targetPaths = [
    {1, 3, 4, 5, 6, 9, 11, 15, 16, 17, 18, 19, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 67, 68, 69, 73, 75, 77,
     82, 83, 84, 86, 87, 99, 102, 103, 104, 105},
    {1, 3, 4, 5, 6, 9, 11, 15, 16, 17, 18, 19, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 70, 71, 75, 77,
     82, 83, 84, 86, 87, 99, 102, 103, 104, 105},
    {1, 3, 4, 5, 6, 11, 15, 16, 17, 18, 19, 21, 22, 23, 25, 35, 36, 50, 52, 64, 65, 66, 68, 69, 70, 71, 73, 75, 77, 82,
     83, 84, 86, 87, 98, 103, 104, 105},
    {1, 3, 4, 5, 6, 11, 15, 16, 17, 19, 21, 22, 23, 24, 25, 26, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 73, 75, 82, 83,
     84, 86, 87, 102, 103, 104, 105},
    {1, 3, 4, 5, 7, 11, 15, 16, 17, 20, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 71, 73, 75, 77, 82, 83,
     84, 86, 87, 98, 103, 104, 105},
    {2, 6, 9, 11, 15, 16, 17, 18, 19, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 70, 71, 75, 77, 82, 83,
     84, 86, 99, 102, 103, 104, 105},
    {2, 7, 11, 15, 16, 17, 20, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 70, 71, 75, 77, 82, 83, 84, 86,
     87, 100, 101, 103, 104, 105},
    {1, 3, 4, 5, 6, 11, 15, 16, 17, 18, 19, 25, 35, 36, 50, 51, 52, 64, 65, 66, 67, 68, 69, 72, 73, 75, 77, 82, 83, 84,
     86, 87, 98, 103, 104, 105},
    {1, 3, 4, 5, 8, 9, 11, 15, 16, 17, 21, 22, 23, 26, 35, 36, 50, 51, 52, 64, 65, 66, 67, 68, 69, 73, 75, 82, 83, 84,
     86, 87, 99, 103, 104, 105},
    {1, 3, 7, 9, 10, 15, 16, 17, 20, 25, 35, 36, 50, 51, 52, 60, 62, 64, 65, 66, 67, 68, 69, 70, 71, 73, 75, 77, 82, 83,
     84, 86, 87, 102},
    {1, 6, 10, 18, 19, 25, 35, 36, 50, 51, 52, 59, 60, 62, 67, 68, 69, 70, 71, 73, 75, 76, 77, 82, 83, 84, 86, 87, 98},
    {6, 11, 15, 16, 17, 19, 21, 22, 23, 24, 25, 26, 35, 36, 50, 52, 63, 64, 65, 66, 68, 69, 75, 82, 83, 84, 86, 102},
    {1, 6, 10, 18, 19, 25, 35, 36, 50, 51, 52, 61, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 82, 83, 84, 86, 87},
    {1, 7, 9, 10, 20, 25, 27, 28, 29, 35, 36, 59, 60, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 102},
    {1, 7, 10, 20, 25, 28, 29, 37, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 91, 92, 93, 100, 101},
    {1, 6, 10, 18, 19, 25, 28, 29, 37, 62, 67, 68, 69, 72, 73, 74, 75, 76, 77, 86, 87, 91, 92, 93, 94, 98},
    {6, 18, 19, 28, 29, 46, 55, 61, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 91, 92, 93},
    {7, 20, 28, 29, 34, 46, 55, 61, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 91, 92, 93},
    {7, 20, 28, 29, 34, 46, 55, 59, 60, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 85, 103, 104},
    {6, 18, 19, 27, 28, 29, 34, 37, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 95, 96},
    {8, 25, 28, 29, 34, 46, 53, 54, 55, 59, 60, 62, 67, 68, 69, 73, 74, 75, 85, 103, 104},
    {7, 12, 13, 14, 20, 32, 33, 44, 46, 48, 49, 53, 54, 55, 70, 71, 76, 77, 85, 97},
    {6, 18, 19, 30, 31, 32, 33, 38, 39, 42, 43, 48, 49, 72, 78, 79, 80, 81, 97},
    {7, 20, 30, 32, 33, 38, 39, 40, 43, 48, 49, 70, 71, 78, 79, 81, 88, 89, 97},
    {6, 18, 19, 32, 33, 45, 48, 49, 53, 54, 55, 70, 71, 76, 77, 85, 97},
    {6, 18, 19, 32, 33, 48, 49, 54, 55, 56, 57, 58, 70, 71, 78, 85, 97},
    {6, 18, 19, 30, 31, 38, 39, 42, 43, 47, 70, 71, 78, 79, 80, 81},
    {6, 19, 26, 32, 33, 41, 48, 49, 79, 97, 100, 101},
    {7, 20, 32, 33, 41, 70, 71, 78, 79, 90, 97},
]

target_paths = [set(path) for path in targetPaths]


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0


def compute_path_similarity_matrix(paths):
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix


def group_paths_by_similarity(paths):
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.mean(avg_sim_scores)

    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]
    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)

    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group


# === random grouping ===
USE_KEYBOARD_INPUT_GROUP_SIZE = False
RANDOM_GROUP_SEED = None


def group_paths_randomly(paths, use_keyboard_input=False, seed=None):
    n_paths = len(paths)
    original_group1, original_group2 = group_paths_by_similarity(paths)
    default_group1_size = len(original_group1)

    group1_size = default_group1_size

    if use_keyboard_input:
        while True:
            user_input = input(
                f"Random group1Number of Paths,  1~{n_paths - 1}; "
                f" {default_group1_size}: "
            ).strip()
            if user_input == "":
                group1_size = default_group1_size
                break
            try:
                group1_size = int(user_input)
                if 1 <= group1_size <= n_paths - 1:
                    break
                print(f": Random group1 1~{n_paths - 1} .")
            except ValueError:
                print(": , .")

    if not (1 <= group1_size <= n_paths - 1):
        raise ValueError(f"Random group1 1~{n_paths - 1} ,  {group1_size}")

    all_indices = list(range(n_paths))
    rng = random.Random(seed) if seed is not None else random
    rng.shuffle(all_indices)

    random_group1 = sorted(all_indices[:group1_size])
    random_group2 = sorted(all_indices[group1_size:])

    return random_group1, random_group2, default_group1_size, len(original_group2)


def compute_robustness(state, path):
    dx, dy, dz = state
    base = execute_Tr(dx, dy, dz)
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dw in [-1, 0, 1]:
        for dt in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dw == dt == dz == 0:
                    continue
                neighbor = np.clip(np.array(state) + np.array([dw, dt, dz]),
                                   STATE_MIN, STATE_MAX)
                neighbor = tuple(neighbor)
                ndx, ndy, ndz = neighbor
                n_trig = execute_Tr(ndx, ndy, ndz)
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def compute_q_value_score(state, similar_model):
    if similar_model is None:
        return 0.0
    try:
        normalized_state = normalizer.normalize(state)
        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = similar_model(state_tensor)
            max_q_value = torch.max(q_values).item()
            normalized_q = max_q_value / 20.0
            normalized_q = max(0.0, min(1.0, normalized_q))
            return 1.0 - normalized_q
    except:
        return 0.0


def generate_samples_for_similar_paths(similar_group, num_candidates=2000, top_k=200, run_id=1):
    SIMILAR_WEIGHTS = [0.55, 0.39, 0.06]

    def save_samples(path_id, samples, base_dir, group_type="similar"):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_{group_type}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"{group_type.title()} Group Path {path_id} - Run {run_id}\n")
            f.write("dx dy dz\tScore\tSimilarity\tLengthDiff\tRobustness\n")
            for s in samples:
                dx, dy, dz = s['state']
                f.write(
                    f"{dx} {dy} {dz}\t{s['score']:.4f}\t{s['similarity']:.4f}\t{s['length_diff']:.4f}\t{s['robustness']:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_grouped"

    for path_idx in similar_group:
        path = target_paths[path_idx]
        path_id = path_idx + 1
        candidate_samples = []
        attempts = 0

        while len(candidate_samples) < num_candidates and attempts < num_candidates * 10:
            attempts += 1
            state = random_state()
            dx, dy, dz = state
            triggered = execute_Tr(dx, dy, dz)
            if not triggered:
                continue
            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            candidate_samples.append({
                'state': state,
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            })

        if candidate_samples:
            for sample in candidate_samples:
                score = (SIMILAR_WEIGHTS[0] * sample['similarity'] +
                         SIMILAR_WEIGHTS[1] * sample['length_diff'] +
                         SIMILAR_WEIGHTS[2] * sample['robustness'])
                sample['score'] = score
            candidate_samples.sort(key=lambda x: x['score'], reverse=True)
            selected_samples = candidate_samples[:top_k]
            save_samples(path_id=path_id, samples=selected_samples, base_dir=base_dir, group_type="similar")


def generate_samples_for_isolated_paths(isolated_group, similar_model, num_candidates=2000, top_k=200, run_id=1):
    ISOLATED_WEIGHTS = [0.18, 0.21, 0.32, 0.29]

    def save_samples(path_id, samples, base_dir, group_type="isolated"):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_{group_type}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"{group_type.title()} Group Path {path_id} - Run {run_id}\n")
            f.write("dx dy dz\tScore\tSimilarity\tLengthDiff\tRobustness\tQValueScore\n")
            for s in samples:
                dx, dy, dz = s['state']
                f.write(
                    f"{dx} {dy} {dz}\t{s['score']:.4f}\t{s['similarity']:.4f}\t{s['length_diff']:.4f}\t{s['robustness']:.4f}\t{s['q_value_score']:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_grouped"

    for path_idx in isolated_group:
        path = target_paths[path_idx]
        path_id = path_idx + 1
        candidate_samples = []
        attempts = 0

        while len(candidate_samples) < num_candidates and attempts < num_candidates * 10:
            attempts += 1
            state = random_state()
            dx, dy, dz = state
            triggered = execute_Tr(dx, dy, dz)
            if not triggered:
                continue
            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            q_score = compute_q_value_score(state, similar_model)
            candidate_samples.append({
                'state': state,
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'q_value_score': q_score,
                'triggered': triggered
            })

        if candidate_samples:
            for sample in candidate_samples:
                score = (ISOLATED_WEIGHTS[0] * sample['similarity'] +
                         ISOLATED_WEIGHTS[1] * sample['length_diff'] +
                         ISOLATED_WEIGHTS[2] * sample['robustness'] +
                         ISOLATED_WEIGHTS[3] * sample['q_value_score'])
                sample['score'] = score
            candidate_samples.sort(key=lambda x: x['score'], reverse=True)
            selected_samples = candidate_samples[:top_k]
            save_samples(path_id=path_id, samples=selected_samples, base_dir=base_dir, group_type="isolated")


class GroupExperienceReplay:
    def __init__(self, capacity=20000):
        self.capacity = capacity
        self.buffer = deque(maxlen=self.capacity)
        self.priorities = deque(maxlen=self.capacity)
        self.sampled_indices = set()

    def append(self, experience):
        self.buffer.append(experience)
        self.priorities.append(experience[-1])

    def sample(self, batch_size, alpha=0.6):
        priorities = np.array(self.priorities) ** alpha
        probabilities = priorities / np.sum(priorities)
        batch_indices = np.random.choice(len(self.buffer), batch_size, p=probabilities)
        batch = [self.buffer[idx] for idx in batch_indices]
        return batch, batch_indices, probabilities[batch_indices]

    def update_priorities(self, batch_indices, td_errors):
        for idx, td_error in zip(batch_indices, td_errors):
            if idx < len(self.priorities):
                self.priorities[idx] = max(td_error, 1e-6)

    def __len__(self):
        return len(self.buffer)

    def get_high_reward_samples(self, target_path, num_samples=20):
        if len(self.buffer) == 0:
            return []

        samples_with_recalculated_scores = []
        for idx, experience in enumerate(self.buffer):
            if idx in self.sampled_indices:
                continue
            normalized_state_tensor = experience[0]
            normalized_state = normalized_state_tensor.cpu().numpy().flatten()
            state_tuple = tuple(normalizer.denormalize(normalized_state))
            dx, dy, dz = state_tuple
            triggered = execute_Tr(dx, dy, dz)
            new_reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_recalculated_scores.append((idx, state_tuple, new_reward, sim, triggered))

        samples_with_recalculated_scores.sort(key=lambda x: x[2], reverse=True)
        selected = samples_with_recalculated_scores[:num_samples]
        for item in selected:
            self.sampled_indices.add(item[0])
        return [(s[1], s[2], s[3], s[4]) for s in selected]

    def reset_sampled_indices(self):
        self.sampled_indices.clear()


def load_path_data(file_path):
    path_data = []
    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines[2:]:
            parts = line.strip().split("\t")
            state = tuple(map(int, parts[0].split()))
            path_data.append(state)
    return path_data


class DQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, 128)
        self.fc2 = nn.Linear(128, 64)
        self.fc3 = nn.Linear(64, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)


class DQNAgentWithPER:
    def __init__(self, state_dim, action_dim, replay_buffer, gamma=0.99, epsilon=1.0, epsilon_decay=0.995,
                 epsilon_min=0.1, learning_rate=0.001, alpha=0.6, beta=0.4):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer
        self.alpha = alpha
        self.beta = beta

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        delta_values = [1, -1]
        dim = action_idx // 2
        delta_idx = action_idx % 2
        delta = delta_values[delta_idx]
        if dim == 0:
            return (delta, 0, 0)
        elif dim == 1:
            return (0, delta, 0)
        elif dim == 2:
            return (0, 0, delta)

    def act(self, normalized_state):
        if random.random() < self.epsilon:
            return random.randrange(self.action_dim)
        state = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state)
        return torch.argmax(q_values, dim=1).item()

    def store_transition(self, normalized_state, action, reward, normalized_next_state, done):
        state = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        next_state = torch.tensor(normalized_next_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state)
            next_q_values = self.target_model(next_state)
            max_next_q_values = next_q_values.max(1)[0]
            target_q_values = reward + (self.gamma * max_next_q_values * (1 - done))
            td_error = torch.abs(q_values[0][action] - target_q_values).item()
        self.replay_buffer.append((state, action, reward, next_state, done, td_error))
        return td_error

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return
        batch, batch_indices, probabilities = self.replay_buffer.sample(batch_size, alpha=self.alpha)
        states, actions, rewards, next_states, dones, _ = zip(*batch)
        weights = (len(self.replay_buffer) * probabilities) ** (-self.beta)
        weights = weights / weights.max()
        weights = torch.tensor(weights, dtype=torch.float32).to(device)

        states = torch.tensor(np.array([s.cpu().numpy().flatten() for s in states]), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array([ns.cpu().numpy().flatten() for ns in next_states]),
                                   dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))
        td_errors = current_q_values - target_q_values
        weighted_loss = (td_errors.pow(2) * weights).mean()

        self.optimizer.zero_grad()
        weighted_loss.backward()
        self.optimizer.step()

        new_priorities = torch.abs(td_errors).detach().cpu().numpy()
        self.replay_buffer.update_priorities(batch_indices, new_priorities)

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())


def train_group(group_paths, path_documents, replay_buffer, batch_size=32, group_name="", pretrained_model=None,
                sample_group_type=None):
    if sample_group_type is None:
        sample_group_type = 'similar' if group_name == '' else 'isolated'
    state_dim = 3
    action_dim = 6

    agent = DQNAgentWithPER(state_dim, action_dim, replay_buffer)

    if pretrained_model is not None:
        print(f"  {group_name}: ()...")
        agent.model.load_state_dict(pretrained_model.state_dict())
        agent.target_model.load_state_dict(pretrained_model.state_dict())
        print(f"  {group_name}: completed")

    path_rewards = {}

    print(f"Start training{group_name}, Included Paths: {[idx + 1 for idx in group_paths]}")
    start_time = time.time()

    BATCH_SIZE = 50
    N_SAMPLES = 200
    N_STEPS = 3
    N_ROUNDS = 5
    N_BATCHES = 4

    replay_count = 0

    for path_idx in group_paths:
        file_path = os.path.join(path_documents,
                                 f"path{path_idx + 1}_{sample_group_type}.txt")
        if not os.path.exists(file_path):
            print(f"    : Path {path_idx + 1}, ")
            continue

        path_data = load_path_data(file_path)
        target_path = target_paths[path_idx]

        if path_idx not in path_rewards:
            path_rewards[path_idx] = 0

        print(f"\n  Start training path  {path_idx + 1},  {N_ROUNDS} ")

        for round_idx in range(N_ROUNDS):
            print(f"    Path  {path_idx + 1} - Run  {round_idx + 1}/{N_ROUNDS} ")

            for batch_idx in range(N_BATCHES):
                batch_start = batch_idx * BATCH_SIZE
                batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)

                if batch_start >= len(path_data):
                    print(f"       {batch_idx + 1}: , ")
                    break

                print(f"       {batch_idx + 1}/{N_BATCHES} ( {batch_start}-{batch_end})")

                for sample_idx in range(batch_start, batch_end):
                    if sample_idx >= len(path_data):
                        break

                    state = path_data[sample_idx]
                    prev_state = None
                    prev_triggered = None

                    for step in range(N_STEPS):
                        normalized_state = normalizer.normalize(state)

                        legal_actions = []
                        for a in range(agent.action_dim):
                            dw, dt, dz = agent.decode_action(a)
                            cand_next = tuple(np.clip(np.array(state) + np.array([dw, dt, dz]),
                                                      STATE_MIN, STATE_MAX))
                            legal_actions.append(a)

                        if not legal_actions:
                            break

                        if random.random() < agent.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        dw, dt, dz = agent.decode_action(action)
                        next_state = tuple(np.clip(np.array(state) + np.array([dw, dt, dz]),
                                                   STATE_MIN, STATE_MAX))

                        normalized_next_state = normalizer.normalize(next_state)

                        dx, dy, dz = next_state
                        triggered = execute_Tr(dx, dy, dz)
                        reward = compute_reward(next_state, target_path, triggered,
                                                prev_triggered, prev_state)
                        done = (step == N_STEPS - 1)

                        agent.store_transition(normalized_state, action, reward, normalized_next_state, done)

                        prev_state = state
                        prev_triggered = triggered
                        state = next_state
                        path_rewards[path_idx] += reward

                if len(agent.replay_buffer) >= batch_size:
                    agent.train(batch_size)
                    replay_count += 1
                    if replay_count % 2 == 0:
                        agent.update_target_model()

            print(f"      Path  {path_idx + 1} - Run  {round_idx + 1} completed")

        print(f"  Path  {path_idx + 1}  {N_ROUNDS} All completed ")

    training_time = time.time() - start_time
    print(f"\n{group_name}completed!")
    print(f"  : Path completed{N_ROUNDS}()")
    print(f"  : {replay_count}")
    print(f"  : {training_time:.2f} seconds")
    print(f"  : {len(replay_buffer)}")

    return agent, path_rewards, training_time


def generate_and_train_grouped_paths_staged(path_documents, random_group1, random_group2, batch_size=32, run_id=1):
    print(f"\n===  {run_id}/20 (3 minutes, random grouping+model reuse) ===")
    random_group1_paths = [idx + 1 for idx in random_group1]
    random_group2_paths = [idx + 1 for idx in random_group2]

    print(f"Random group1Path (pretrained group): {random_group1_paths}")
    print(f"Random group2Path (model-reuse group): {random_group2_paths}")

    total_start_time = time.time()

    print(f"\n[1] Random group1...")
    generate_samples_for_similar_paths(random_group1, num_candidates=2000, top_k=200, run_id=run_id)

    print(f"\n[2] Random group1(, {5})...")
    group1_replay_buffer = GroupExperienceReplay(capacity=20000)
    group1_agent, group1_path_rewards, group1_training_time = train_group(
        random_group1, path_documents, group1_replay_buffer, batch_size,
        group_name="Random group1(pretrained group)", pretrained_model=None, sample_group_type="similar"
    )

    print(f"\n[3] Random group1Random group2...")
    generate_samples_for_isolated_paths(random_group2, group1_agent.model,
                                        num_candidates=2000, top_k=200, run_id=run_id)

    print(f"\n[4] Random group2(Random group1, {5})...")
    group2_replay_buffer = GroupExperienceReplay(capacity=20000)
    group2_agent, group2_path_rewards, group2_training_time = train_group(
        random_group2, path_documents, group2_replay_buffer, batch_size,
        group_name="Random group2(model-reuse group)", pretrained_model=group1_agent.model, sample_group_type="isolated"
    )

    total_path_rewards = {**group1_path_rewards, **group2_path_rewards}
    total_cumulative_reward = sum(total_path_rewards.values())
    total_training_time = time.time() - total_start_time

    print(f"\n===  {run_id}/20 completed, : {total_training_time:.2f} seconds ===")
    print(f"Random group1: {group1_training_time:.2f} seconds")
    print(f"Random group2: {group2_training_time:.2f} seconds")
    print(f" - Random group1: {len(group1_replay_buffer)}, Random group2: {len(group2_replay_buffer)}")

    return group1_agent, group2_agent, group1_replay_buffer, group2_replay_buffer, \
        total_cumulative_reward, total_path_rewards, total_training_time


def create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir):
    os.makedirs(output_dir, exist_ok=True)

    similar_group_paths = [idx + 1 for idx in similar_group]
    isolated_group_paths = [idx + 1 for idx in isolated_group]

    wb = Workbook()

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    header_color = "4472C4"
    similar_group_color = "E2EFDA"
    isolated_group_color = "FCE4D6"
    stats_color = "FFF2CC"

    ws_paths = wb.active
    ws_paths.title = "Path "

    path_headers = ['Path ID', ''] + [f'Run {i}' for i in range(1, 21)] + ['Average Similarity', 'Maximum Similarity',
                                                                                    'Minimum Similarity', 'Standard deviation']
    for col, header in enumerate(path_headers, 1):
        cell = ws_paths.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_paths.row_dimensions[1].height = 30

    for path_id in range(1, len(target_paths) + 1):
        row = path_id + 1

        if path_id in similar_group_paths:
            group_type = "Random group1(pretrained group)"
            row_color = similar_group_color
        elif path_id in isolated_group_paths:
            group_type = "Random group2(model-reuse group)"
            row_color = isolated_group_color
        else:
            group_type = "Ungrouped"
            row_color = "FFFFFF"

        cell = ws_paths.cell(row=row, column=1, value=f"Path {path_id}")
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        cell.border = thin_border

        cell = ws_paths.cell(row=row, column=2, value=group_type)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        cell.border = thin_border

        path_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            sim = run_data['path_similarities'].get(path_id, {}).get('avg_similarity', 0.0)
            path_similarities.append(sim)

            cell = ws_paths.cell(row=row, column=3 + run_idx, value=round(sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        stats_values = [
            round(np.mean(path_similarities), 4),
            round(np.max(path_similarities), 4),
            round(np.min(path_similarities), 4),
            round(np.std(path_similarities), 4)
        ]

        for i, value in enumerate(stats_values):
            cell = ws_paths.cell(row=row, column=23 + i, value=value)
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
            cell.font = Font(bold=True, size=10)
            cell.border = thin_border

    ws_paths.column_dimensions['A'].width = 13
    ws_paths.column_dimensions['B'].width = 16
    for col in range(3, 23):
        ws_paths.column_dimensions[get_column_letter(col)].width = 10
    for col in range(23, 27):
        ws_paths.column_dimensions[get_column_letter(col)].width = 13

    ws_groups = wb.create_sheet("")

    group_headers = ['Group Name', 'Included Paths'] + [f'Run {i}' for i in range(1, 21)] + ['Average Similarity', 'Standard deviation']
    for col, header in enumerate(group_headers, 1):
        cell = ws_groups.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_groups.row_dimensions[1].height = 30

    row = 2

    cell = ws_groups.cell(row=row, column=1, value="Random group1(pretrained group)")
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    cell.border = thin_border

    cell = ws_groups.cell(row=row, column=2, value=','.join(map(str, similar_group_paths)))
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    cell.border = thin_border

    group_similarities = []
    for run_idx, run_data in enumerate(all_runs_data):
        group_sim = np.mean(
            [run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0) for p in similar_group_paths])
        group_similarities.append(group_sim)

        cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(group_sim, 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    cell = ws_groups.cell(row=row, column=23, value=round(np.mean(group_similarities), 4))
    cell.number_format = '0.0000'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border

    cell = ws_groups.cell(row=row, column=24, value=round(np.std(group_similarities), 4))
    cell.number_format = '0.0000'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border

    row += 1

    if isolated_group_paths:
        cell = ws_groups.cell(row=row, column=1, value="Random group2(model-reuse group)")
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        cell.border = thin_border

        cell = ws_groups.cell(row=row, column=2, value=','.join(map(str, isolated_group_paths)))
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        cell.border = thin_border

        isolated_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            iso_sim = np.mean(
                [run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0) for p in isolated_group_paths])
            isolated_similarities.append(iso_sim)

            cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(iso_sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        cell = ws_groups.cell(row=row, column=23, value=round(np.mean(isolated_similarities), 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border

        cell = ws_groups.cell(row=row, column=24, value=round(np.std(isolated_similarities), 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border

    ws_groups.column_dimensions['A'].width = 16
    ws_groups.column_dimensions['B'].width = 22
    for col in range(3, 23):
        ws_groups.column_dimensions[get_column_letter(col)].width = 10
    ws_groups.column_dimensions[get_column_letter(23)].width = 14
    ws_groups.column_dimensions[get_column_letter(24)].width = 12

    ws_samples = wb.create_sheet("Detailed Sample Data")

    sample_headers = ['Run', 'Path ID', 'Sample ID', 'Dx', 'Dy', 'Dz', 'Similarity', 'Triggered Rule Set']
    for col, header in enumerate(sample_headers, 1):
        cell = ws_samples.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_samples.row_dimensions[1].height = 30

    sample_row = 2
    for run_idx, run_data in enumerate(all_runs_data, 1):
        for path_id in range(1, len(target_paths) + 1):
            samples = run_data['path_samples'].get(path_id, [])

            if path_id in similar_group_paths:
                path_color = similar_group_color
            elif path_id in isolated_group_paths:
                path_color = isolated_group_color
            else:
                path_color = "FFFFFF"

            for sample_idx, (state_tuple, reward, sim, triggered) in enumerate(samples, 1):
                dx, dy, dz = state_tuple
                triggered_str = ','.join(map(str, sorted(triggered)))

                cell = ws_samples.cell(row=sample_row, column=1, value=f"Run {run_idx}")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=path_color, end_color=path_color, fill_type="solid")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=2, value=f"Path {path_id}")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=path_color, end_color=path_color, fill_type="solid")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=3, value=sample_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                for col_offset, value in enumerate([dx, dy, dz]):
                    cell = ws_samples.cell(row=sample_row, column=4 + col_offset, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=7, value=round(sim, 4))
                cell.number_format = '0.0000'
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=8, value=f"{{{triggered_str}}}")
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border

                sample_row += 1

    sample_widths = [13, 13, 11, 10, 12, 8, 12, 45]
    for i, width in enumerate(sample_widths, 1):
        ws_samples.column_dimensions[get_column_letter(i)].width = width

    output_path = os.path.join(output_dir, "20 run_random grouping_model reuse_3 minutes.xlsx")
    wb.save(output_path)
    print(f"\n Consolidated Excel report generated: {output_path}")


def run_20_times_training():
    model_path_base = r"D:\Experiment\CNN\DQNNEW\saved_models_random_reuse_3min_version"
    path_documents = r"D:\Experiment\CNN\DQNNEW\path_samples_grouped"
    output_dir = r"D:\Experiment\CNN\ComparisonExperiment2\excel_reports_random_reuse_3min_version"

    os.makedirs(model_path_base, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    similar_group, isolated_group, default_group1_size, default_group2_size = group_paths_randomly(
        target_paths,
        use_keyboard_input=USE_KEYBOARD_INPUT_GROUP_SIZE,
        seed=RANDOM_GROUP_SEED
    )

    similar_group_display = [idx + 1 for idx in similar_group]
    isolated_group_display = [idx + 1 for idx in isolated_group]

    print("=" * 60)
    print("20 - random grouping + model reuse - 3 minutes")
    print("=" * 60)
    print("Training-scale configuration:")
    print("   Per path: 5")
    print("   Per round: 4")
    print("   Per batch: 50")
    print("   Per sample: 3")
    print("   Sample generation: 2000candidates -> 200final samples")
    print("   : save model parameters only(optimized version)")
    print("   : ")
    print(f"   Default group size: Random group1={default_group1_size}Path , Random group2={default_group2_size}Path ")
    print(f"   Keyboard-input group size: {'' if USE_KEYBOARD_INPUT_GROUP_SIZE else ''}")
    print(f"   Random seed: {RANDOM_GROUP_SEED if RANDOM_GROUP_SEED is not None else 'None,  run'}")
    print("=" * 60)
    print(f"\nAutomatic grouping results:")
    print(f"Similar path group: {similar_group_display}")
    print(f"Isolated path group: {isolated_group_display}")
    print("\n" + "=" * 60)

    all_runs_data = []
    total_start_time = time.time()

    for run_id in range(1, 21):
        print(f"\n{'=' * 60}")
        print(f"Start run  {run_id}/20  run")
        print(f"{'=' * 60}")

        group1_agent, group2_agent, group1_buffer, group2_buffer, total_cumulative_reward, path_rewards, training_time = \
            generate_and_train_grouped_paths_staged(path_documents, similar_group, isolated_group, batch_size=32,
                                                    run_id=run_id)

        group1_model_path = os.path.join(model_path_base, f"random_group1_model_run_{run_id}.pth")
        group2_model_path = os.path.join(model_path_base, f"random_group2_model_run_{run_id}.pth")

        torch.save(group1_agent.model.state_dict(), group1_model_path)
        torch.save(group2_agent.model.state_dict(), group2_model_path)

        print(f"[Run {run_id}] Model saved(optimized version - )")

        group1_buffer.reset_sampled_indices()
        group2_buffer.reset_sampled_indices()

        run_data = {
            'run_id': run_id,
            'training_time': training_time,
            'total_reward': total_cumulative_reward,
            'path_rewards': path_rewards,
            'path_similarities': {},
            'path_samples': {}
        }

        all_similarities = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            path_id = path_idx + 1

            if path_id in similar_group_display:
                buffer = group1_buffer
            elif path_id in isolated_group_display:
                buffer = group2_buffer
            else:
                continue

            high_reward_samples = buffer.get_high_reward_samples(target_path, num_samples=20)

            if high_reward_samples:
                similarities = [sim for _, _, sim, _ in high_reward_samples]
                run_data['path_similarities'][path_idx + 1] = {
                    'avg_similarity': np.mean(similarities),
                    'max_similarity': np.max(similarities),
                    'min_similarity': np.min(similarities),
                    'sample_count': len(similarities)
                }
                run_data['path_samples'][path_idx + 1] = high_reward_samples
                all_similarities.extend(similarities)
            else:
                run_data['path_similarities'][path_idx + 1] = {
                    'avg_similarity': 0.0,
                    'max_similarity': 0.0,
                    'min_similarity': 0.0,
                    'sample_count': 0
                }
                run_data['path_samples'][path_idx + 1] = []

        if all_similarities:
            run_data['overall_avg_similarity'] = np.mean(all_similarities)
            run_data['max_similarity'] = np.max(all_similarities)
            run_data['min_similarity'] = np.min(all_similarities)
        else:
            run_data['overall_avg_similarity'] = 0.0
            run_data['max_similarity'] = 0.0
            run_data['min_similarity'] = 0.0

        all_runs_data.append(run_data)

        print(f"[Run {run_id}] completed! Overall Average Similarity: {run_data['overall_avg_similarity']:.4f}")
        print(f"{'=' * 60}\n")

    total_time = time.time() - total_start_time

    print("\nGenerating consolidated Excel report...")
    create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir)

    print("\n" + "=" * 60)
    print("20All completed! - random grouping + model reuse - 3 minutes")
    print("=" * 60)
    print(f":")
    print(f"  Per path: 5 x 4 x 50 x 3 = 3000/Path ")
    print(f"  Sample generation: 2000candidates -> 200final samples")
    print(f"  : save model parameters only(optimized version)")
    print(f"  Total elapsed time: {total_time:.2f} seconds ({total_time / 60:.2f} minutes)")
    print(f"  Average elapsed time per run: {total_time / 20:.2f} seconds")
    print(f"\nAverage similarity statistics:")
    avg_similarities = [r['overall_avg_similarity'] for r in all_runs_data]
    print(f"  Overall average: {np.mean(avg_similarities):.4f}")
    print(f"  Maximum: {np.max(avg_similarities):.4f}")
    print(f"  Minimum: {np.min(avg_similarities):.4f}")
    print(f"  Standard deviation: {np.std(avg_similarities):.4f}")
    print(f"\nAll results have been saved to: {output_dir}")
    print("=" * 60)


if __name__ == "__main__":
    run_20_times_training()