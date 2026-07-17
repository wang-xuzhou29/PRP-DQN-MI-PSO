import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
from torch.distributions import Normal
import numpy as np
import random
import time
import math
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# === 配置（新范围：温度 1~300，电压 1~200，流量 1~5） ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'ACTION_DIM': 3,
    'MIN_VALUES': np.array([1, 1, 1], dtype=np.float32),
    'MAX_VALUES': np.array([300, 200, 5], dtype=np.float32),
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 3,
    'REPLAY_BATCH_SIZE': 64,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'ACTOR_LR': 3e-4,
    'CRITIC_LR': 3e-4,
    'ALPHA_LR': 3e-4,
    'GAMMA': 0.99,
    'TAU': 0.005,
    'NUM_RUNS': 20,
    'TOP_K_SAMPLES': 20,
    'TARGET_PATHS': [
        {1, 3, 4, 5, 6, 9, 11, 15, 16, 17, 18, 19, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 67, 68, 69, 73, 75, 77,
         82, 83, 84, 86, 87, 99, 102, 103, 104, 105},
        {1, 3, 4, 5, 6, 9, 11, 15, 16, 17, 18, 19, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 70, 71, 75, 77,
         82, 83, 84, 86, 87, 99, 102, 103, 104, 105},
        {1, 3, 4, 5, 6, 11, 15, 16, 17, 18, 19, 21, 22, 23, 25, 35, 36, 50, 52, 64, 65, 66, 68, 69, 70, 71, 73, 75, 77, 82,
         83, 84, 86, 87, 98, 103, 104, 105},
        {1, 3, 4, 5, 6, 11, 15, 16, 17, 19, 21, 22, 23, 24, 25, 26, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 73, 75, 82, 83,
         84, 86, 87, 102, 103, 104, 105},
        {1, 3, 4, 5, 7, 11, 15, 16, 17, 20, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 71, 73, 75, 77, 82, 83,
         84, 86, 87, 98, 103, 104, 105},
        {2, 6, 9, 11, 15, 16, 17, 18, 19, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 70, 71, 75, 77, 82, 83,
         84, 86, 99, 102, 103, 104, 105},
        {2, 7, 11, 15, 16, 17, 20, 21, 22, 23, 25, 35, 36, 50, 51, 52, 64, 65, 66, 68, 69, 70, 71, 75, 77, 82, 83, 84, 86,
         87, 100, 101, 103, 104, 105},
        {1, 3, 4, 5, 6, 11, 15, 16, 17, 18, 19, 25, 35, 36, 50, 51, 52, 64, 65, 66, 67, 68, 69, 72, 73, 75, 77, 82, 83, 84,
         86, 87, 98, 103, 104, 105},
        {1, 3, 4, 5, 8, 9, 11, 15, 16, 17, 21, 22, 23, 26, 35, 36, 50, 51, 52, 64, 65, 66, 67, 68, 69, 73, 75, 82, 83, 84,
         86, 87, 99, 103, 104, 105},
        {1, 3, 7, 9, 10, 15, 16, 17, 20, 25, 35, 36, 50, 51, 52, 60, 62, 64, 65, 66, 67, 68, 69, 70, 71, 73, 75, 77, 82, 83,
         84, 86, 87, 102},
        {1, 6, 10, 18, 19, 25, 35, 36, 50, 51, 52, 59, 60, 62, 67, 68, 69, 70, 71, 73, 75, 76, 77, 82, 83, 84, 86, 87, 98},
        {6, 11, 15, 16, 17, 19, 21, 22, 23, 24, 25, 26, 35, 36, 50, 52, 63, 64, 65, 66, 68, 69, 75, 82, 83, 84, 86, 102},
        {1, 6, 10, 18, 19, 25, 35, 36, 50, 51, 52, 61, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 82, 83, 84, 86, 87},
        {1, 7, 9, 10, 20, 25, 27, 28, 29, 35, 36, 59, 60, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 102},
        {1, 7, 10, 20, 25, 28, 29, 37, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 91, 92, 93, 100, 101},
        {1, 6, 10, 18, 19, 25, 28, 29, 37, 62, 67, 68, 69, 72, 73, 74, 75, 76, 77, 86, 87, 91, 92, 93, 94, 98},
        {6, 18, 19, 28, 29, 46, 55, 61, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 91, 92, 93},
        {7, 20, 28, 29, 34, 46, 55, 61, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 91, 92, 93},
        {7, 20, 28, 29, 34, 46, 55, 59, 60, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 85, 103, 104},
        {6, 18, 19, 27, 28, 29, 34, 37, 62, 67, 68, 69, 70, 71, 73, 74, 75, 76, 77, 86, 87, 95, 96},
        {8, 25, 28, 29, 34, 46, 53, 54, 55, 59, 60, 62, 67, 68, 69, 73, 74, 75, 85, 103, 104},
        {7, 12, 13, 14, 20, 32, 33, 44, 46, 48, 49, 53, 54, 55, 70, 71, 76, 77, 85, 97},
        {6, 18, 19, 30, 31, 32, 33, 38, 39, 42, 43, 48, 49, 72, 78, 79, 80, 81, 97},
        {7, 20, 30, 32, 33, 38, 39, 40, 43, 48, 49, 70, 71, 78, 79, 81, 88, 89, 97},
        {6, 18, 19, 32, 33, 45, 48, 49, 53, 54, 55, 70, 71, 76, 77, 85, 97},
        {6, 18, 19, 32, 33, 48, 49, 54, 55, 56, 57, 58, 70, 71, 78, 85, 97},
        {6, 18, 19, 30, 31, 38, 39, 42, 43, 47, 70, 71, 78, 79, 80, 81},
        {6, 19, 26, 32, 33, 41, 48, 49, 79, 97, 100, 101},
        {7, 20, 32, 33, 41, 70, 71, 78, 79, 90, 97},
    ]
}

# === section2 触发函数（规则编号 1~105，与 targetPaths 匹配） ===
def section2_flow_pressure_density_hybrid(x, y, z):
    b = {}
    triggered = set()

    if (30 < x < 170) != (30 < x < 100): b[0] = 1
    if (30 < x < 170) != (30 < x < 190): b[1] = 2
    if (30 < x < 170) != (30 < x < 140): b[2] = 3
    if (30 < x < 170) != (30 < x < 150): b[3] = 4
    if (30 < x < 170) != (30 < x < 160): b[4] = 5
    if (100 < y < 200) != (150 < y < 200): b[5] = 6
    if (100 < y < 200) != (50 < y < 200): b[6] = 7
    if (100 < y < 200) != (100 < y < 150): b[7] = 8
    if (1 < z < 4) != (1 < z < 3): b[8] = 9
    if (x > 10 and x < 152) != (x > 10 and x < 100): b[9] = 10

    if (x > 40 and x < 152) != (x > 40 and x < 202): b[10] = 11
    if (abs(x - 50) < 4) != (abs(x - 50) < 2): b[11] = 12
    if (abs(x - 50) < 4) != (abs(x + 50) < 4): b[12] = 13
    if (abs(x - 50) < 4) != (abs(x * 50) < 4): b[13] = 14
    if (x > 135) != (x > 215): b[14] = 15
    if (x > 135) != (x > 235): b[15] = 16
    if (x > 135) != (x > 225): b[16] = 17
    if (y > 100) != (y > 120): b[17] = 18
    if (y > 100) != (y > 140): b[18] = 19
    if (y > 100) != (y > 50): b[19] = 20

    if (x < 165) != (x < 265): b[20] = 21
    if (x < 165) != (x < 245): b[21] = 22
    if (x < 165) != (x < 262): b[22] = 23
    if (y < 150) != (y < 130): b[23] = 24
    if (y < 150) != (x < 100): b[24] = 25
    if (y < 190) != (y < 120): b[25] = 26
    if ((x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 125) != (
            (x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 115): b[26] = 27
    if ((x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 125) != (
            (x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 85): b[27] = 28
    if ((x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 125) != (
            (x + y / 30 + z / 2) > 95 and (x + y / 30 + z / 2) < 25): b[28] = 29
    if (x / (y / 30) > 1 and x / (y / 30) < 2) != (x / (y / 30) > 1 and x / (y / 30) < 4): b[29] = 30

    if (x / (y / 30) > 1 and x / (y / 30) < 2) != (x / (y / 30) > 1 and x / (y / 30) < 3): b[30] = 31
    if (x / (z / 2) > 2 and x / (z / 2) < 4) != (x / (z / 2) > 1 and x / (z / 2) < 4): b[31] = 32
    if (x / (z / 2) > 2 and x / (z / 2) < 4) != (x / (z / 3) > 2 and x / (z / 2) < 4): b[32] = 33
    if (x / (z / 2) > 2 and x / (z / 2) < 4) != (x / (z / 2) > 2 and x / (z / 3) < 4): b[33] = 34
    if ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 55) != ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 155): b[34] = 35
    if ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 55) != ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 125): b[35] = 36
    if ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 55) != ((x + y / 30) / 2 > 45 and (x + y / 30) / 2 < 25): b[36] = 37
    if (x - y / 30 > -10 and x - y / 30 < 10) != (x - y / 30 > -10 and x - y / 30 < 5): b[37] = 38
    if (x - y / 30 > -10 and x - y / 30 < 10) != (x - y / 30 > -10 and x - y / 30 < 6): b[38] = 39
    if (x - y / 30 > -10 and x - y / 30 < 10) != (x - y / 30 > -10 and x - y / 30 < 8): b[39] = 40

    if (abs(x / (y / 30) - 1.0) < 1) != (abs(x / (y / 30) + 1.0) < 1): b[40] = 41
    if (abs(x / (y / 30) - 1.0) < 1) != (abs(x / (y / 30) - 1.0) < 2): b[41] = 42
    if (abs(x / (y / 30) - 1.0) < 1) != (abs(x / (y / 30) - 1.0) < 3): b[42] = 43
    if (x / ((y / 30) + 20) > 1 and x / ((y / 30) + 20) < 2) != (x / ((y / 30) + 20) > 1 and x / ((y / 20) + 20) < 2): b[43] = 44
    if (x / ((y / 30) + 20) > 1 and x / ((y / 30) + 20) < 2) != (x / ((y / 30) + 20) > 1 and x / ((y / 40) + 20) < 2): b[44] = 45
    if (x / ((y / 30) + 20) > 1 and x / ((y / 30) + 20) < 2) != (x / ((y / 30) + 20) > 1 and x / ((y / 30) + 40) < 2): b[45] = 46
    if ((z / 50) / (x / 50) > 1 and (z / 50) / (x / 50) < 3) != ((z / 50) / (x / 50) > 1 and (z / 50) / (x / 50) < 4): b[46] = 47
    if ((z / 50) / (x / 50) > 1 and (z / 50) / (x / 50) < 3) != ((z / 50) / (x / 50) > 2 and (z / 50) / (x / 50) < 3): b[47] = 48
    if ((z / 50) / (x / 50) > 1 and (z / 50) / (x / 50) < 3) != ((z / 50) / (x / 50) > 3 and (z / 50) / (x / 50) < 3): b[48] = 49
    if ((x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 45) != (
            (x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 70): b[49] = 50

    if ((x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 45) != (
            (x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 60): b[50] = 51
    if ((x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 3 < 45) != (
            (x + y / 30 + z / 2) / 3 > 35 and (x + y / 30 + z / 2) / 6 < 45): b[51] = 52
    if (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 8) != (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 48): b[52] = 53
    if (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 8) != (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 50): b[53] = 54
    if (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 8) != (math.sqrt((x - 50) ** 2 + (y / 30 - 50) ** 2) < 58): b[54] = 55
    if ((x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 200) != (
            (x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 400): b[55] = 56
    if ((x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 200) != (
            (x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 300): b[56] = 57
    if ((x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 200) != (
            (x - 40) * (y / 30 - 40) > 50 and (x - 40) * (y / 30 - 40) < 500): b[57] = 58
    if (abs((x + y / 30) - 100) < 15) != (abs((x + y / 30) - 100) < 35): b[58] = 59
    if (abs((x + y / 30) - 100) < 15) != (abs((x + y / 30) - 100) < 45): b[59] = 60

    if (abs((x + y / 30) - 100) < 15) != (abs((x + y / 30) - 100) < 10): b[60] = 61
    if (x / 50 > 1 and x / 50 < 4) != (x / 50 > 3 and x / 50 < 4): b[61] = 62
    if (x / 50 > 1 and x / 50 < 4) != (x / 50 > 1 and x / 50 < 5): b[62] = 63
    if (x < 135) != (x < 215): b[63] = 64
    if (x < 135) != (x < 225): b[64] = 65
    if (x < 135) != (x < 225): b[65] = 66
    if (x > 65) != (x > 165): b[66] = 67
    if (x > 65) != (x > 215): b[67] = 68
    if (x > 65) != (x > 251): b[68] = 69
    if (y < 110) != (y < 80): b[69] = 70

    if (y < 110) != (y < 50): b[70] = 71
    if (y < 110) != (y < 119): b[71] = 72
    if (x < 32 or x > 68) != (x < 32 or x > 168): b[72] = 73
    if (x < 32 or x > 68) != (x < 32 or x > 118): b[73] = 74
    if (x < 32 or x > 68) != (x < 32 or x > 218): b[74] = 75
    if (x < 35 and y < 120) != (x < 135 and y < 120): b[75] = 76
    if (x < 35 and y < 120) != (x < 235 and y < 120): b[76] = 77
    if (x < 35 and y < 120) != (x < 35 and y < 60): b[77] = 78
    if (abs(x / (y / 30) - 1.0) > 1) != (abs(x / (y / 30) - 3.0) > 1): b[78] = 79
    if (abs(x / (y / 30) - 1.0) > 1) != (abs(x / (y / 30) - 1.0) > 2): b[79] = 80

    if (abs(x / (y / 30) - 1.0) > 1) != (abs(x / (y / 30) - 1.0) > 3): b[80] = 81
    if ((x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 135) != (
            (x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 235): b[81] = 82
    if ((x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 135) != (
            (x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 335): b[82] = 83
    if ((x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 135) != (
            (x + y / 30 + z / 2) < 85 or (x + y / 30 + z / 2) > 435): b[83] = 84
    if (x < 30 or x > 70) != (x < 30 or x * 7 > 70): b[84] = 85
    if (x < 30 or x > 70) != (x < 30 or x > 270): b[85] = 86
    if (x < 30 or x > 70) != (x < 30 or x > 170): b[86] = 87
    if (x / (y / 30) > 1 and x / (y / 30) < 3) != (x / (y / 30) > 1 and x / (y / 20) < 3): b[87] = 88
    if (x / (y / 30) > 1 and x / (y / 30) < 3) != (x / (y / 30) > 1 and x / (y / 30) < 4): b[88] = 89
    if (x / (y / 30) > 1 and x / (y / 30) < 3) != (x / (y / 20) > 1 and x / (y / 30) < 3): b[89] = 90

    if ((x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 112) != (
            (x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 62): b[90] = 91
    if ((x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 112) != (
            (x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 92): b[91] = 92
    if ((x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 112) != (
            (x + y / 30 + z / 2) > 108 and (x + y / 30 + z / 2) < 82): b[92] = 93
    if (abs((x + y / 30) / 2 - 50) < 1) != (abs((x + y / 30) / 2 - 50) < 4): b[93] = 94
    if (abs((x + y / 30) / 2 - 50) < 1) != (abs((x + y / 30) / 2 + 50) < 1): b[94] = 95
    if (abs((x + y / 30) / 2 - 50) < 1) != (abs((x + y / 20) / 2 + 50) < 1): b[95] = 96
    if (x / (z / 2) > 2 and x / (z / 2) < 4) != (x / (z / 2) > 1 and x / (z / 2) < 4): b[96] = 97
    if (x / (z / 2) > 20 and x / (z / 2) < 100) != (x / (z / 3) > 20 and x / (z / 2) < 100): b[97] = 98
    if (x / (z / 2) > 20 and x / (z / 2) < 100) != (x / (z / 2) > 20 and x / (z / 2) < 150): b[98] = 99
    if ((y / 30) / (z / 50) > 38 and (y / 30) / (z / 50) < 42) != (
            (y / 30) / (z / 50) > 48 and (y / 30) / (z / 50) < 42): b[99] = 100

    if ((y / 30) / (z / 50) > 38 and (y / 30) / (z / 50) < 42) != (
            (y / 30) / (z / 50) > 58 and (y / 30) / (z / 50) < 42): b[100] = 101
    if ((y / 30) / (z / 50) > 38 and (y / 30) / (z / 50) < 42) != (
            (y / 30) / (z / 50) > 38 and (y / 30) / (z / 50) < 72): b[101] = 102
    if ((x - 40) / 20 > 1 and (x - 40) / 20 < 6) != ((x - 40) / 30 > 1 and (x - 40) / 20 < 7): b[102] = 103
    if ((x - 40) / 20 > 1 and (x - 40) / 20 < 6) != ((x - 50) / 20 > 1 and (x - 40) / 20 < 7): b[103] = 104
    if ((x - 40) / 20 > 1 and (x - 40) / 20 < 6) != ((x - 40) / 20 > 1 and (x - 50) / 20 < 7): b[104] = 105

    triggered = set(b.values())
    return triggered

# === 绑定执行函数（使用 section2） ===
execute_Tr = section2_flow_pressure_density_hybrid

# === 状态处理辅助函数 ===
def clip_state(state):
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return np.clip(state, min_vals, max_vals)

def normalize_state(state):
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return 2 * (state - min_vals) / (max_vals - min_vals) - 1

def denormalize_state(normalized_state):
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return (normalized_state + 1) * (max_vals - min_vals) / 2 + min_vals

def coverage_similarity(triggered, target_path):
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0
    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)

def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']
    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']
    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']
    return reward

# === SAC Actor ===
class GaussianPolicy(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dim=256):
        super(GaussianPolicy, self).__init__()
        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.mean = nn.Linear(hidden_dim, action_dim)
        self.log_std = nn.Linear(hidden_dim, action_dim)
        self.action_scale = 8.0
        self.action_bias = 0.0

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))
        mean = self.mean(x)
        log_std = self.log_std(x)
        log_std = torch.clamp(log_std, min=-20, max=2)
        return mean, log_std

    def sample(self, state):
        mean, log_std = self.forward(state)
        std = log_std.exp()
        normal = Normal(mean, std)
        x_t = normal.rsample()
        action = torch.tanh(x_t) * self.action_scale + self.action_bias
        log_prob = normal.log_prob(x_t)
        log_prob -= torch.log(self.action_scale * (1 - torch.tanh(x_t).pow(2)) + 1e-6)
        log_prob = log_prob.sum(1, keepdim=True)
        mean = torch.tanh(mean) * self.action_scale + self.action_bias
        return action, log_prob, mean

# === SAC Critic ===
class QNetwork(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dim=256):
        super(QNetwork, self).__init__()
        self.fc1 = nn.Linear(state_dim + action_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, 1)
        self.fc4 = nn.Linear(state_dim + action_dim, hidden_dim)
        self.fc5 = nn.Linear(hidden_dim, hidden_dim)
        self.fc6 = nn.Linear(hidden_dim, 1)

    def forward(self, state, action):
        xu = torch.cat([state, action], 1)
        x1 = F.relu(self.fc1(xu))
        x1 = F.relu(self.fc2(x1))
        q1 = self.fc3(x1)
        x2 = F.relu(self.fc4(xu))
        x2 = F.relu(self.fc5(x2))
        q2 = self.fc6(x2)
        return q1, q2

# === 经验回放缓冲区（包含路径和相似度信息） ===
class EnhancedReplayBuffer:
    def __init__(self, capacity=50000):
        self.buffer = deque(maxlen=capacity)
        self.experience_info = deque(maxlen=capacity)

    def push(self, state, action, reward, next_state, done, path_idx, similarity):
        self.buffer.append((state, action, reward, next_state, done))
        self.experience_info.append({'path_idx': path_idx, 'similarity': similarity})

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return None
        batch = random.sample(self.buffer, batch_size)
        state, action, reward, next_state, done = map(np.stack, zip(*batch))
        return (
            torch.FloatTensor(state).to(device),
            torch.FloatTensor(action).to(device),
            torch.FloatTensor(reward).unsqueeze(1).to(device),
            torch.FloatTensor(next_state).to(device),
            torch.FloatTensor(done).unsqueeze(1).to(device)
        )

    def get_top_k_per_path(self, num_paths, k=20):
        path_samples = {i: [] for i in range(num_paths)}
        for idx, info in enumerate(self.experience_info):
            path_idx = info['path_idx']
            similarity = info['similarity']
            path_samples[path_idx].append((idx, similarity, self.buffer[idx]))

        top_k_results = {}
        for path_idx in range(num_paths):
            samples = path_samples[path_idx]
            if len(samples) == 0:
                top_k_results[path_idx] = []
                continue
            samples.sort(key=lambda x: x[1], reverse=True)
            top_k = samples[:k]
            top_k_results[path_idx] = []
            for sample in top_k:
                normalized_state = sample[2][0]
                original_state = denormalize_state(normalized_state)
                original_state_rounded = np.round(original_state).astype(int)
                x, y, z = original_state_rounded
                triggered = execute_Tr(x, y, z)
                top_k_results[path_idx].append({
                    'state': original_state_rounded,
                    'similarity': sample[1],
                    'triggered': triggered
                })
        return top_k_results

    def __len__(self):
        return len(self.buffer)

# === SAC Agent ===
class SACAgent:
    def __init__(self, state_dim=3, action_dim=3):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.policy = GaussianPolicy(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.policy_optimizer = optim.Adam(self.policy.parameters(), lr=EXPERIMENT_CONFIG['ACTOR_LR'])
        self.critic = QNetwork(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_target = QNetwork(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_target.load_state_dict(self.critic.state_dict())
        self.critic_optimizer = optim.Adam(self.critic.parameters(), lr=EXPERIMENT_CONFIG['CRITIC_LR'])
        self.target_entropy = -action_dim
        self.log_alpha = torch.zeros(1, requires_grad=True, device=device)
        self.alpha_optimizer = optim.Adam([self.log_alpha], lr=EXPERIMENT_CONFIG['ALPHA_LR'])
        self.replay_buffer = EnhancedReplayBuffer()
        self.replay_train_count = 0

    def get_action(self, state, deterministic=False):
        normalized_state = normalize_state(state)
        state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)
        with torch.no_grad():
            if deterministic:
                _, _, action = self.policy.sample(state_tensor)
            else:
                action, _, _ = self.policy.sample(state_tensor)
        action = action.cpu().numpy()[0]
        return action

    def store_experience(self, state, action, reward, next_state, done, path_idx, similarity):
        normalized_state = normalize_state(state)
        normalized_next_state = normalize_state(next_state)
        self.replay_buffer.push(
            normalized_state, action, reward,
            normalized_next_state, done, path_idx, similarity
        )

    def replay_train(self):
        batch = self.replay_buffer.sample(EXPERIMENT_CONFIG['REPLAY_BATCH_SIZE'])
        if batch is None:
            return
        state, action, reward, next_state, done = batch

        with torch.no_grad():
            next_action, next_log_prob, _ = self.policy.sample(next_state)
            q1_next, q2_next = self.critic_target(next_state, next_action)
            q_next = torch.min(q1_next, q2_next)
            target_q = reward + (1 - done) * EXPERIMENT_CONFIG['GAMMA'] * (
                    q_next - self.log_alpha.exp() * next_log_prob)

        q1, q2 = self.critic(state, action)
        critic_loss = F.mse_loss(q1, target_q) + F.mse_loss(q2, target_q)

        self.critic_optimizer.zero_grad()
        critic_loss.backward()
        self.critic_optimizer.step()

        new_action, log_prob, _ = self.policy.sample(state)
        q1_new, q2_new = self.critic(state, new_action)
        q_new = torch.min(q1_new, q2_new)
        policy_loss = (self.log_alpha.exp() * log_prob - q_new).mean()

        self.policy_optimizer.zero_grad()
        policy_loss.backward()
        self.policy_optimizer.step()

        alpha_loss = -(self.log_alpha * (log_prob + self.target_entropy).detach()).mean()
        self.alpha_optimizer.zero_grad()
        alpha_loss.backward()
        self.alpha_optimizer.step()

        for param, target_param in zip(self.critic.parameters(), self.critic_target.parameters()):
            target_param.data.copy_(
                EXPERIMENT_CONFIG['TAU'] * param.data + (1 - EXPERIMENT_CONFIG['TAU']) * target_param.data)

        self.replay_train_count += 1
        if self.replay_train_count % 2 == 0:
            alpha_value = self.log_alpha.exp().item()
            print(f"  -> SAC 更新 (第 {self.replay_train_count} 次), Alpha={alpha_value:.4f}")

# === 性能计算函数 ===
def calculate_run_performance(run_idx, sac_results, training_time, total_steps, update_count, agent):
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)
    all_similarities = []
    total_samples = 0
    all_rewards = []
    total_reward = 0

    for path_idx in range(num_paths):
        samples = sac_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']
            total_reward += reward
            all_rewards.append(reward)
            all_similarities.append(similarity)
            total_samples += 1

    average_reward = total_reward / total_samples if total_samples > 0 else 0
    convergence = np.mean(all_similarities) if all_similarities else 0
    environment_adaptability = 1 / (np.std(all_similarities) + 1e-8) if len(all_similarities) > 1 else 0
    generalization_ability = convergence
    computational_efficiency = total_steps / training_time if training_time > 0 else 0
    policy_update_frequency = update_count / training_time if training_time > 0 else 0
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0

    return {
        'Run': run_idx + 1,
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }

# === Excel 导出函数 ===
def export_to_excel(all_sac_results, all_performance_data, target_paths, output_path="SAC_20_run.xlsx"):
    print("\n正在导出数据到 Excel...")
    all_sac_summary_data = []
    all_sac_detailed_data = []

    for run_idx, (sac_results, performance_data) in enumerate(zip(all_sac_results, all_performance_data)):
        sac_summary_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = sac_results[path_idx]
            if len(samples) == 0:
                sac_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue
            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = 'Yes' if perfect_count > 0 else 'No'
            sac_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })
        all_sac_summary_data.extend(sac_summary_data)

        sac_detailed_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = sac_results[path_idx]
            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']
                sac_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Sample ID': sample_idx + 1,
                    'X(温度)': int(state[0]),
                    'Y(电压)': int(state[1]),
                    'Z(流量)': int(state[2]),
                    'Similarity': round(similarity, 4),
                    'Perfect Coverage': 'Yes' if abs(similarity - 1.0) < 0.001 else 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path))),
                    'Triggered Rules': ', '.join(map(str, sorted(triggered))),
                    'Hit Rule Count': len(target_path.intersection(triggered)),
                    'Target Rule Count': len(target_path)
                })
        all_sac_detailed_data.extend(sac_detailed_data)

    sac_summary_df = pd.DataFrame(all_sac_summary_data)
    sac_detailed_df = pd.DataFrame(all_sac_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        sac_summary_df.to_excel(writer, sheet_name='SACPath', index=False)
        sac_detailed_df.to_excel(writer, sheet_name='SACDetailed Sample Data', index=False)
        selected_columns = [
            'Run', 'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df[selected_columns].to_excel(writer, sheet_name='Metric', index=False)

        workbook = writer.book
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        perfect_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        ws1 = writer.sheets['SACPath']
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row_idx in range(2, ws1.max_row + 1):
            if ws1.cell(row_idx, 9).value == 'Yes':
                for col_idx in range(1, ws1.max_column + 1):
                    ws1.cell(row_idx, col_idx).fill = perfect_fill
        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 18
        ws1.column_dimensions['D'].width = 15
        ws1.column_dimensions['E'].width = 18
        ws1.column_dimensions['F'].width = 15
        ws1.column_dimensions['G'].width = 15
        ws1.column_dimensions['H'].width = 15
        ws1.column_dimensions['I'].width = 18
        ws1.column_dimensions['J'].width = 50

        ws2 = writer.sheets['SACDetailed Sample Data']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 18
        ws2.column_dimensions['I'].width = 40
        ws2.column_dimensions['J'].width = 40
        ws2.column_dimensions['K'].width = 18
        ws2.column_dimensions['L'].width = 18

        ws3 = writer.sheets['Metric']
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws3.column_dimensions[col].width = 20

    print(f"文件已保存到: {output_path}")
    print(f"  - Sheet1: SACPath 统计 ({len(all_sac_summary_data)} 条)")
    print(f"  - Sheet2: SACDetailed Sample Data ({len(all_sac_detailed_data)} 条)")
    print(f"  - Sheet3: Metric ({len(all_performance_data)} 条)")

# === 单次训练流程 ===
def train_sac_workflow():
    print("=" * 80)
    print("SAC 训练 - 流量压力密度混合规则集 (section2)")
    print("状态范围: X(温度) 1~300, Y(电压) 1~200, Z(流量) 1~5")
    print("相似度 = 命中规则数 / 目标路径规则数")
    print("=" * 80)

    agent = SACAgent()
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n生成初始样本: {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']} / 路径")
    path_samples = {}
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']

    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            state = np.random.uniform(min_vals, max_vals).astype(np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  路径 {path_idx + 1}/{num_paths}: {len(samples)} 个初始样本")

    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size

    print(f"\n批次大小: {batch_size}, 单样本迭代步数: {EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(f"总批次数: {num_batches} /路径 × {num_paths} 路径 = {num_batches * num_paths}")
    print("-" * 80)

    for batch_idx in range(num_batches):
        print(f"\n批次 {batch_idx + 1}/{num_batches}")

        for path_idx in range(num_paths):
            target_path = target_paths[path_idx]
            batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]

            batch_rewards = []
            batch_similarities = []

            for initial_state in batch_samples:
                state = initial_state.copy()
                episode_reward = 0
                final_similarity = 0

                for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                    action = agent.get_action(state, deterministic=False)
                    next_state = state + action
                    next_state = clip_state(next_state)

                    x, y, z = next_state
                    triggered = execute_Tr(x, y, z)
                    reward = unified_reward_function(triggered, target_path)
                    similarity = coverage_similarity(triggered, target_path)

                    done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)

                    agent.store_experience(
                        state, action, reward, next_state, done,
                        path_idx, similarity
                    )

                    state = next_state
                    episode_reward += reward
                    final_similarity = similarity
                    total_steps += 1

                batch_rewards.append(episode_reward)
                batch_similarities.append(final_similarity)

            avg_reward = np.mean(batch_rewards)
            avg_similarity = np.mean(batch_similarities)
            print(f"  路径 {path_idx + 1}: 平均奖励={avg_reward:.2f}, 平均相似度={avg_similarity:.4f}")

        print(f"\n  执行 SAC 更新...")
        agent.replay_train()
        print(f"  经验池大小: {len(agent.replay_buffer)}")

    training_time = time.time() - start_time

    print("\n" + "=" * 80)
    print(f"单次运行训练完成! 总耗时: {training_time:.2f} 秒, 总交互步数: {total_steps}")
    print(f"经验池大小: {len(agent.replay_buffer)}")
    print(f"SAC 更新次数: {agent.replay_train_count}")
    print("=" * 80)

    print(f"\n提取每条路径的 Top-{EXPERIMENT_CONFIG['TOP_K_SAMPLES']} 高相似度样本...")
    top_k_results = agent.replay_buffer.get_top_k_per_path(num_paths, EXPERIMENT_CONFIG['TOP_K_SAMPLES'])

    return agent, top_k_results, training_time, total_steps, agent.replay_train_count

# === 主程序 ===
def main():
    print("\n" + "=" * 80)
    print("SAC 多轮实验 - 流量压力密度混合规则测试 (section2)")
    print("参数范围: 温度 1~300, 电压 1~200, 流量 1~5")
    print("=" * 80)

    all_sac_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'='*80}")
        print(f"开始第 {run_idx + 1}/{EXPERIMENT_CONFIG['NUM_RUNS']} 次独立运行")
        print(f"{'='*80}")

        sac_agent, sac_results, training_time, total_steps, update_count = train_sac_workflow()
        performance_data = calculate_run_performance(
            run_idx, sac_results, training_time, total_steps, update_count, sac_agent
        )

        all_sac_results.append(sac_results)
        all_performance_data.append(performance_data)

        print(f"\n第 {run_idx + 1} 次运行完成!")
        print(f"  总奖励: {performance_data['Total Reward']}")
        print(f"  平均奖励: {performance_data['Average Reward']}")
        print(f"  平均相似度: {performance_data['Average Similarity']}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"SAC_flow_pressure_density_20run_{timestamp}.xlsx"
    export_to_excel(all_sac_results, all_performance_data, target_paths, output_path)

    print("\n" + "=" * 80)
    print(f"{EXPERIMENT_CONFIG['NUM_RUNS']} 次运行整体统计摘要")
    print("=" * 80)

    total_rewards = [p['Total Reward'] for p in all_performance_data]
    average_rewards = [p['Average Reward'] for p in all_performance_data]
    convergences = [p['Convergence'] for p in all_performance_data]
    environment_adaptabilities = [p['Environment Adaptability'] for p in all_performance_data]
    generalization_abilities = [p['Generalization Ability'] for p in all_performance_data]
    computational_efficiencies = [p['Computational Efficiency'] for p in all_performance_data]
    policy_update_frequencies = [p['Policy Update Frequency'] for p in all_performance_data]
    avg_similarities = [p['Average Similarity'] for p in all_performance_data]

    print(f"\nTotal Reward Statistics:")
    print(f"  Mean: {np.mean(total_rewards):.2f}")
    print(f"  Std: {np.std(total_rewards):.2f}")

    print(f"\nAverage Reward Statistics:")
    print(f"  Mean: {np.mean(average_rewards):.4f}")
    print(f"  Std: {np.std(average_rewards):.4f}")

    print(f"\nConvergence Statistics:")
    print(f"  Mean: {np.mean(convergences):.4f}")
    print(f"  Std: {np.std(convergences):.4f}")

    print(f"\nEnvironment Adaptability Statistics:")
    print(f"  Mean: {np.mean(environment_adaptabilities):.4f}")
    print(f"  Std: {np.std(environment_adaptabilities):.4f}")

    print(f"\nGeneralization Ability Statistics:")
    print(f"  Mean: {np.mean(generalization_abilities):.4f}")
    print(f"  Std: {np.std(generalization_abilities):.4f}")

    print(f"\nComputational Efficiency Statistics:")
    print(f"  Mean: {np.mean(computational_efficiencies):.2f}")
    print(f"  Std: {np.std(computational_efficiencies):.2f}")

    print(f"\nPolicy Update Frequency Statistics:")
    print(f"  Mean: {np.mean(policy_update_frequencies):.4f}")
    print(f"  Std: {np.std(policy_update_frequencies):.4f}")

    print("\n" + "=" * 80)
    print(f" {EXPERIMENT_CONFIG['NUM_RUNS']} 次运行全部完成!")
    print("=" * 80)

if __name__ == "__main__":
    main()