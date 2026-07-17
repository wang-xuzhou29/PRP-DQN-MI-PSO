import os
import random
import math
import numpy as np
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# --- 状态范围配置（按您的需求）---
STATE_MIN_X, STATE_MAX_X = 1, 100   # 温度
STATE_MIN_Y, STATE_MAX_Y = 1, 100   # 电压
STATE_MIN_Z, STATE_MAX_Z = 1, 60    # 流量

def generate_input():
    return [
        random.randint(STATE_MIN_X, STATE_MAX_X),
        random.randint(STATE_MIN_Y, STATE_MAX_Y),
        random.randint(STATE_MIN_Z, STATE_MAX_Z)
    ]

# ========== 规则函数（section4） ==========
def section4_comprehensive_hybrid_control(x, y, z):
    # 确保数值在合理范围内（外部范围已限定，但保留作为安全防护）
    x = max(STATE_MIN_X, min(STATE_MAX_X, x))
    y = max(STATE_MIN_Y, min(STATE_MAX_Y, y))
    z = max(STATE_MIN_Z, min(STATE_MAX_Z, z))

    triggered = set()

    if [(30 < x < 70)] != [(30 < x * 200 < 70)]: triggered.add(1)
    if [(30 < x < 70)] != [(30 < 6000 < 70)]: triggered.add(2)
    if [(30 < y < 70)] != [(30 < 7000 < 70)]: triggered.add(3)
    if [(30 < y < 70)] != [(30 < y * 600 < 70)]: triggered.add(4)
    if [(10 < z < 50)] != [(10 < 9000 < 50)]: triggered.add(5)
    if [(10 < z < 50)] != [(10 < z * 66 < 50)]: triggered.add(6)
    if [(x > 40 and x < 60 and y > 30 and y < 70)] != [(x > 40 and x < 60 and y > 30 and 9 < 70)]: triggered.add(7)
    if [(x > 35 and x < 65 and z > 20 and z < 40)] != [(x > 35 and x < 565 and z > 20 and z < 40)]: triggered.add(8)
    if [(x > 35 and x < 65 and z > 20 and z < 40)] != [(x > 35 and x * 11 < 65 and z > 20 and z < 40)]: triggered.add(9)
    if [(y > 30 and y < 70 and z > 20 and z < 40)] != [(y > 30 and y * 78 < 70 and z > 20 and z < 40)]: triggered.add(10)
    if [(y > 30 and y < 70 and z > 20 and z < 40)] != [(y > 30 and y < 70 and z * 56 > 20 and z < 40)]: triggered.add(11)
    if [(x > 45 and x < 55)] != [(x > 45 and x < 155)]: triggered.add(12)
    if [(y > 45 and y < 55)] != [(y > 45 and y < 88)]: triggered.add(13)
    if [(z > 25 and z < 35)] != [(z > 25 and z < 88)]: triggered.add(14)
    if [(z > 25 and z < 35)] != [(z > 288 and z < 35)]: triggered.add(15)
    if [(abs(x - 50) < 5)] != [(abs(x - 50) < 55)]: triggered.add(16)
    if [(abs(x - 50) < 5)] != [(abs(x - 50) < 533)]: triggered.add(17)
    if [(abs(y - 50) < 5)] != [(abs(y - 50) < 522)]: triggered.add(18)
    if [(abs(y - 50) < 5)] != [(abs(y - 50) < 95)]: triggered.add(19)
    if [(abs(z - 30) < 3)] != [(abs(z - 30) < 89)]: triggered.add(20)
    if [(abs(z - 30) < 3)] != [(abs(z - 30) < 63)]: triggered.add(21)
    if [(x > 30 + 5)] != [(x * 89 > 30 + 5)]: triggered.add(22)
    if [(x > 30 + 5)] != [(900 > 30 + 5)]: triggered.add(23)
    if [(y > 30 + 5)] != [(789 > 30 + 5)]: triggered.add(24)
    if [(y > 30 + 5)] != [(y * 78 > 30 + 5)]: triggered.add(25)
    if [(z > 10 + 5)] != [(999 > 10 + 5)]: triggered.add(26)
    if [(z > 10 + 5)] != [(z * 9 > 10 + 5)]: triggered.add(27)
    if [(x < 70 - 5)] != [(x * 88 < 70 - 5)]: triggered.add(28)
    if [(x < 70 - 5)] != [(988 < 70 - 5)]: triggered.add(29)
    if [(y < 70 - 5)] != [(y * 789 < 70 - 5)]: triggered.add(30)
    if [(y < 70 - 5)] != [(888 < 70 - 5)]: triggered.add(31)
    if [(z < 50 - 5)] != [(z * 96 < 50 - 5)]: triggered.add(32)
    if [(z < 50 - 5)] != [(777 < 50 - 5)]: triggered.add(33)
    if [(x > 35 and x < 65 and y > 35 and y < 65 and z > 15 and z < 45)] != [
        (x > 35 and x < 651 and y > 35 and y < 65 and z > 15 and z < 45)]: triggered.add(34)
    if [((x / 2 + y / 2 + z / 2) > 80 and (x / 2 + y / 2 + z / 2) < 120)] != [
        ((x / 2 + y / 2 + z / 2) > 810 and (x / 2 + y / 2 + z / 2) < 120)]: triggered.add(35)
    if [((x / 2 + y / 2 + z / 2) > 80 and (x / 2 + y / 2 + z / 2) < 120)] != [
        ((x / 2 + y / 2 + z / 2) > 1000 and (x / 2 + y / 2 + z / 2) < 120)]: triggered.add(36)
    if [(x / y > 0.8 and x / y < 1.2)] != [(x / y > 0.8 and x / y < 12)]: triggered.add(37)
    if [(x / z > 1.0 and x / z < 2.0)] != [(x / z > 1.0 and x / z < 20)]: triggered.add(38)
    if [(y / z > 1.0 and y / z < 2.0)] != [(y / z > 18 and y / z < 2.0)]: triggered.add(39)
    if [(y / z > 1.0 and y / z < 2.0)] != [(y / z > 1.0 and y / z < 20)]: triggered.add(40)
    if [(x * y > 1500 and x * y < 3500)] != [(x * y > 1500 and x * y < 350)]: triggered.add(41)
    if [(x * y > 1500 and x * y < 3500)] != [(x * y > 1500 and x * y < 35)]: triggered.add(42)
    if [(x * z > 500 and x * z < 2000)] != [(x * z > 5100 and x * z < 2000)]: triggered.add(43)
    if [(x * z > 500 and x * z < 2000)] != [(x * z > 5100 and x * z < 2000)]: triggered.add(44)
    if [(y * z > 500 and y * z < 2000)] != [(y * z > 5001 and y * z < 2000)]: triggered.add(45)
    if [(y * z > 500 and y * z < 2000)] != [(y * z > 500 and y * z < 200)]: triggered.add(46)
    if [((x + y) / 2 > 40 and (x + y) / 2 < 60)] != [((x + y) / 2 > 140 and (x + y) / 2 < 60)]: triggered.add(47)
    if [((x + y) / 2 > 40 and (x + y) / 2 < 60)] != [((x + y) / 2 > 40 and (x + y) / 2 < 6110)]: triggered.add(48)
    if [(x - y > -10 and x - y < 10)] != [(x - y > -10 and x - y < 110)]: triggered.add(49)
    if [(x - y > -10 and x - y < 10)] != [(x - y > -10 and x - y < 1220)]: triggered.add(50)
    if [(abs(x / y - 1.0) < 0.2)] != [(abs(x / y - 1.0) < 12)]: triggered.add(51)
    if [(abs(x / y - 1.0) < 0.2)] != [(abs(x / y - 1.0) < 211)]: triggered.add(52)
    if [(x / (y + 10) > 0.7 and x / (y + 10) < 1.3)] != [(x / (y + 10) > 711 and x / (y + 10) < 1.3)]: triggered.add(53)
    if [(x / (y + 10) > 0.7 and x / (y + 10) < 1.3)] != [(x / (y + 10) > 0.7 and x / (y + 10) < 13)]: triggered.add(54)
    if [(y / (z + 5) > 0.8 and y / (z + 5) < 1.5)] != [(y / (z + 5) > 0.8 and y / (z + 5) < 1225)]: triggered.add(55)
    if [(y / (z + 5) > 0.8 and y / (z + 5) < 1.5)] != [(y / (z + 5) > 0.8 and y / (z + 5) < 15)]: triggered.add(56)
    if [(z / (x + 5) > 0.5 and z / (x + 5) < 1.2)] != [(z / (x + 5) > 5 and z / (x + 5) < 1.2)]: triggered.add(57)
    if [(z / (x + 5) > 0.5 and z / (x + 5) < 1.2)] != [(z / (x + 5) > 511 and z / (x + 5) < 1.2)]: triggered.add(58)
    if [((x + y + z) / 3 > 30 and (x + y + z) / 3 < 50)] != [
        ((x + y + z) / 3 > 130 and (x + y + z) / 3 < 50)]: triggered.add(59)
    if [(math.sqrt((x - 50) ** 2 + (y - 50) ** 2) < 10)] != [
        (math.sqrt((x - 50) ** 2 + (y - 50) ** 2) < 1110)]: triggered.add(60)
    if [((x - 45) * (y - 45) > 100 and (x - 45) * (y - 45) < 1000)] != [
        ((x - 45) * (y - 45) > 100 and (x - 45) * (y - 45) < 200)]: triggered.add(61)
    if [((x - 45) * (y - 45) > 100 and (x - 45) * (y - 45) < 1000)] != [
        ((x - 45) * (y - 45) > 100 and (x - 45) * (y - 45) < 10)]: triggered.add(62)
    if [(abs((x + y) - 100) < 20)] != [(abs((x + y) - 100) < 201)]: triggered.add(63)
    if [(abs((x + y) - 100) < 20)] != [(abs((x + y) - 100) < 2110)]: triggered.add(64)
    if [(x / 50 > 0.8 and x / 50 < 1.2)] != [(x / 50 > 118 and x / 50 < 1.2)]: triggered.add(65)
    if [(x < 30)] != [(x < 220)]: triggered.add(66)
    if [(x > 70)] != [(x > 30)]: triggered.add(67)
    if [(x > 70)] != [(x > 440)]: triggered.add(68)
    if [(y < 30)] != [(y < 302)]: triggered.add(69)
    if [(y < 30)] != [(y < 130)]: triggered.add(70)
    if [(y > 70)] != [(y > 170)]: triggered.add(71)
    if [(y > 70)] != [(y > 7)]: triggered.add(72)
    if [(z < 15)] != [(z < 30)]: triggered.add(73)
    if [(z > 45)] != [(z > 450)]: triggered.add(74)
    if [(z > 45)] != [(z > 645)]: triggered.add(75)
    if [(x < 25 or x > 75)] != [(x < 25 or x > 575)]: triggered.add(76)
    if [(x < 25 or x > 75)] != [(x < 25 or x > 715)]: triggered.add(77)
    if [(y < 25 or y > 75)] != [(y < 25 or y > 751)]: triggered.add(78)
    if [(x > 65 and y > 65)] != [(x > 65 and y > 965)]: triggered.add(79)
    if [(x > 65 and y > 65)] != [(x > 65 and y > 635)]: triggered.add(80)
    if [(x < 35 and z < 20)] != [(x < 353 and z < 20)]: triggered.add(81)
    if [(x > 65 and z > 40)] != [(x > 65 and z > 140)]: triggered.add(82)
    if [(x > 65 and z > 40)] != [(x > 65 and z > 410)]: triggered.add(83)
    if [(y > 65 and z > 40)] != [(y > 651 and z > 40)]: triggered.add(84)
    if [(y > 65 and z > 40)] != [(y > 65 and z > 140)]: triggered.add(85)
    if [(abs(x / y - 1.0) > 0.4)] != [(abs(x / y - 1.0) > 4)]: triggered.add(86)
    if [(abs(x / y - 1.0) > 0.4)] != [(abs(x / y - 1.0) > 14)]: triggered.add(87)
    if [((x / 2 + y / 2 + z / 2) < 70 or (x / 2 + y / 2 + z / 2) > 130)] != [
        ((x / 2 + y / 2 + z / 2) < 7333 or (x / 2 + y / 2 + z / 2) > 130)]: triggered.add(88)
    if [((x / 2 + y / 2 + z / 2) < 70 or (x / 2 + y / 2 + z / 2) > 130)] != [
        ((x / 2 + y / 2 + z / 2) < 788 or (x / 2 + y / 2 + z / 2) > 130)]: triggered.add(89)
    if [(x < 20 or x > 80)] != [(x < 120 or x > 80)]: triggered.add(90)
    if [(x < 20 or x > 80)] != [(x < 20 or x > 99)]: triggered.add(91)
    if [(z > 27 and z < 33)] != [(z > 27 and z < 133)]: triggered.add(92)
    if [(z > 27 and z < 33)] != [(z > 27 and z < 331)]: triggered.add(93)
    if [(abs(z - 30) < 1)] != [(abs(z - 30) < 61)]: triggered.add(94)
    if [(abs(z - 30) < 1)] != [(abs(z - 30) < 133)]: triggered.add(95)
    if [(x / y > 0.95 and x / y < 1.05)] != [(x / y > 0.95 and x / y < 105)]: triggered.add(96)
    if [((x / 2 + y / 2 + z / 2) > 90 and (x / 2 + y / 2 + z / 2) < 110)] != [
        ((x / 2 + y / 2 + z / 2) > 999 and (x / 2 + y / 2 + z / 2) < 110)]: triggered.add(97)
    if [(abs((x + y) / 2 - 50) < 2)] != [(abs((x + y) / 2 - 50) < 112)]: triggered.add(98)
    if [(abs((x + y) / 2 - 50) < 2)] != [(abs((x + y) / 2 - 50) < 222)]: triggered.add(99)
    if [(x / z > 1.2 and x / z < 1.8)] != [(x / z > 1.2 and x / z < 18)]: triggered.add(100)
    if [(x / z > 1.2 and x / z < 1.8)] != [(x / z > 1.2 and x / z < 188)]: triggered.add(101)
    if [(y / z > 1.2 and y / z < 1.8)] != [(y / z > 12 and y / z < 1.8)]: triggered.add(102)
    if [(y / z > 1.2 and y / z < 1.8)] != [(y / z > 122 and y / z < 1.8)]: triggered.add(103)
    if [((y - 45) / 10 > 0.45 and (y - 45) / 10 < 0.55)] != [
        ((y - 45) / 10 > 0.45 and (y - 45) / 10 < 755)]: triggered.add(104)
    if [((y - 45) / 10 > 0.45 and (y - 45) / 10 < 0.55)] != [
        ((y - 45) / 10 > 0.45 and (y - 45) / 10 < 55)]: triggered.add(105)
    if [(min(z - 15, 45 - z) > 5)] != [(min(z - 15, 45 - z) > 511)]: triggered.add(106)
    if [(min(z - 15, 45 - z) > 5)] != [(min(z - 15, 45 - z) > 5220)]: triggered.add(107)
    if [(abs(x * y - 2500) < 50)] != [(abs(x * y - 2500) < 5012)]: triggered.add(108)

    return triggered

# ========== 关键修正：别名指向 section4 ==========
execute_Tr = section4_comprehensive_hybrid_control

# ========== Jaccard 相似度 ==========
def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0

# === 目标路径组（编号对应 section4 的规则） ===
targetPaths = [
    {3, 4, 5, 6, 8, 10, 12, 13, 14, 16, 17, 18, 19, 20, 21, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 45, 46, 48, 49, 50,
     51, 52, 54, 60, 61, 62, 63, 64, 66, 68, 69, 70, 72, 76, 77, 86, 87, 88, 89, 91, 92, 93, 94, 95, 96, 97, 98, 99,
     100, 101, 102, 103, 104, 105, 106, 107, 108},

    {3, 4, 5, 6, 8, 10, 12, 13, 14, 16, 17, 18, 19, 20, 21, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 48, 49, 50, 51, 52,
     53, 55, 56, 60, 61, 62, 63, 64, 66, 68, 69, 70, 72, 76, 77, 86, 87, 88, 89, 91, 92, 93, 94, 95, 96, 97, 98, 99,
     100, 101, 102, 103, 104, 105, 106, 107, 108},

    {3, 4, 5, 6, 8, 10, 12, 13, 15, 16, 17, 18, 19, 20, 21, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 45, 46, 48, 49, 50,
     51, 52, 53, 60, 61, 62, 63, 64, 66, 68, 69, 70, 72, 76, 77, 86, 87, 88, 89, 90, 92, 93, 94, 95, 96, 98, 99, 100,
     101, 102, 103, 104, 105, 106, 107, 108},

    {3, 4, 5, 6, 8, 10, 12, 13, 15, 16, 17, 18, 19, 20, 21, 30, 31, 32, 33, 34, 35, 36, 37, 38, 40, 45, 46, 48, 49, 50,
     51, 52, 54, 55, 56, 60, 61, 62, 63, 64, 66, 68, 69, 70, 72, 73, 76, 77, 86, 87, 88, 89, 91, 94, 95, 96, 98, 99,
     100, 101, 104, 105, 106, 107, 108},

    {3, 4, 5, 6, 11, 12, 13, 16, 17, 18, 19, 20, 21, 26, 27, 30, 31, 32, 33, 35, 36, 37, 38, 40, 43, 44, 45, 46, 48, 49,
     50, 51, 52, 54, 55, 56, 60, 61, 62, 63, 64, 66, 68, 69, 70, 72, 73, 76, 77, 81, 86, 87, 88, 89, 91, 94, 95, 96, 98,
     99, 100, 101, 104, 105, 108},

    {3, 4, 5, 6, 8, 10, 12, 13, 14, 16, 17, 18, 19, 20, 21, 32, 33, 35, 36, 37, 38, 39, 48, 49, 50, 51, 52, 53, 55, 56,
     60, 61, 62, 63, 64, 66, 68, 69, 70, 72, 76, 77, 79, 80, 88, 89, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100, 101, 102,
     103, 104, 105, 106, 107, 108},

    {3, 4, 5, 6, 12, 13, 14, 16, 17, 18, 19, 20, 21, 32, 33, 35, 36, 37, 38, 39, 48, 49, 50, 51, 52, 53, 55, 56, 60, 61,
     62, 63, 64, 66, 68, 69, 70, 72, 76, 77, 79, 80, 82, 83, 84, 85, 88, 89, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100,
     101, 102, 103, 104, 105, 108},

    {3, 4, 5, 6, 12, 13, 14, 16, 17, 18, 19, 20, 21, 30, 31, 32, 33, 34, 35, 36, 37, 39, 48, 49, 50, 51, 52, 53, 57, 58,
     60, 61, 62, 63, 64, 66, 68, 69, 70, 72, 76, 77, 82, 83, 86, 87, 88, 89, 90, 92, 93, 94, 95, 96, 98, 99, 100, 101,
     102, 103, 104, 105, 108},

    {3, 4, 5, 6, 12, 13, 14, 16, 17, 18, 19, 20, 21, 35, 36, 37, 39, 48, 49, 50, 51, 52, 53, 57, 58, 60, 61, 62, 63, 64,
     66, 68, 69, 70, 72, 74, 75, 76, 77, 79, 80, 82, 83, 84, 85, 88, 89, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100, 101,
     102, 103, 104, 105, 108},

    {3, 4, 5, 6, 8, 10, 12, 16, 17, 18, 19, 20, 21, 24, 25, 30, 31, 32, 33, 37, 38, 39, 41, 42, 43, 44, 45, 46, 48, 49,
     50, 51, 52, 54, 59, 60, 63, 64, 66, 68, 69, 70, 72, 73, 76, 77, 86, 87, 88, 89, 91, 94, 95, 96, 98, 99, 100, 101,
     102, 103, 106, 107, 108},

    {5, 6, 12, 13, 14, 16, 17, 18, 19, 20, 21, 32, 33, 35, 36, 37, 38, 39, 48, 49, 50, 51, 52, 53, 55, 56, 60, 63, 64,
     66, 68, 69, 70, 71, 76, 77, 78, 79, 80, 82, 83, 84, 85, 88, 89, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100, 101, 102,
     103, 104, 105, 108},

    {1, 2, 3, 4, 5, 6, 9, 10, 12, 16, 17, 18, 19, 20, 21, 24, 25, 28, 29, 30, 31, 32, 33, 37, 38, 39, 41, 42, 43, 44,
     45, 46, 47, 49, 50, 51, 52, 53, 59, 60, 65, 66, 67, 69, 70, 72, 73, 86, 87, 90, 94, 95, 96, 98, 99, 100, 101, 102,
     103, 106, 107, 108},

    {1, 2, 5, 6, 7, 9, 12, 15, 16, 17, 18, 19, 20, 21, 28, 29, 32, 33, 35, 36, 38, 40, 43, 44, 48, 51, 52, 55, 56, 60,
     61, 62, 63, 64, 65, 66, 67, 69, 70, 71, 73, 78, 86, 87, 88, 89, 90, 94, 95, 98, 99, 100, 101, 104, 105, 106, 107,
     108},

    {1, 2, 3, 4, 5, 6, 10, 13, 14, 16, 17, 18, 19, 20, 21, 22, 23, 28, 29, 30, 31, 32, 33, 39, 41, 42, 43, 44, 47, 51,
     52, 55, 56, 57, 58, 59, 60, 66, 67, 69, 70, 72, 86, 87, 90, 92, 93, 94, 95, 98, 99, 102, 103, 104, 105, 106, 107,
     108},
]

# ========== 实验配置 ==========
class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 2000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(targetPaths)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }

def compute_robustness(state, path):
    base = execute_Tr(*state)
    if not base:
        return 0.0
    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                    np.clip(state[0] + dx, STATE_MIN_X, STATE_MAX_X),
                    np.clip(state[1] + dy, STATE_MIN_Y, STATE_MAX_Y),
                    np.clip(state[2] + dz, STATE_MIN_Z, STATE_MAX_Z)
                ])
                n_trig = execute_Tr(*neighbor)
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0

def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = targetPaths[target_path_idx]
    samples = []
    attempts = 0
    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(STATE_MIN_X, STATE_MAX_X),
            random.randint(STATE_MIN_Y, STATE_MAX_Y),
            random.randint(STATE_MIN_Z, STATE_MAX_Z)
        ])
        triggered = execute_Tr(*state)
        if not triggered:
            continue
        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)
        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }
        samples.append(sample_data)
    return samples

def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample['robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)
        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]
    return selected_samples

def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = targetPaths[target_path_idx]
        samples = []
        attempts = 0
        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(STATE_MIN_X, STATE_MAX_X),
                random.randint(STATE_MIN_Y, STATE_MAX_Y),
                random.randint(STATE_MIN_Z, STATE_MAX_Z)
            ])
            triggered = execute_Tr(*state)
            if not triggered:
                continue
            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)
            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)
        random.shuffle(samples)
        return samples[:config.top_k_samples]
    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")
        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)

def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}
    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates
    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}
        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config, shared_candidates[path_idx])
            strategy_results[path_idx] = samples
        results[strategy_name] = strategy_results
    return results

def analyze_fitness_values(results, config):
    analysis_results = {}
    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []
        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])
        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }
        all_scores = []
        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])
        scores_array = np.array(all_scores)
        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)
        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })
        analysis_results[strategy_name] = analysis
    return analysis_results

def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []
    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)
    df = pd.DataFrame(df_data)
    return df, analysis_results

def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []
    print(f"Starting {num_runs} experiments...")
    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")
        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)
        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)
        print(f"Experiment {run_idx + 1} completed.")
    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df

def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "Raw Data"
    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')
    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width
    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []
    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]
        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)
    stats_df = pd.DataFrame(stats_data)
    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)
    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width
    wb.save(output_path)
    print(f"Results saved to: {output_path}")

def main():
    results_df = run_multiple_experiments(num_runs=20)
    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")
    save_results_to_excel(results_df, output_path)
    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)
    return results_df, output_path

if __name__ == "__main__":
    results, output_file = main()