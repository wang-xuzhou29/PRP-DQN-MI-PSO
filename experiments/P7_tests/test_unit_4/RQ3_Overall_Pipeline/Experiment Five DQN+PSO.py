import os
import sys
import random
import time
from collections import deque
from datetime import datetime
from typing import Any, Dict, List, Optional, Set, Tuple

import numpy as np
import torch
import torch.nn as nn
import torch.optim as optim
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ========================================
# ========== experiment configuration parameters ==========
# ========================================
NUM_RUNS = 20

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === three-dimensional range settings ===
# Keep the current DQN state range used by the second script. To use a 0-500 range, modify this section only.
LIGHT_MIN = 1
LIGHT_MAX = 100
MOISTURE_MIN = 1
MOISTURE_MAX = 100
TEMP_MIN = 1
TEMP_MAX = 60
BOUNDS = {
    "light": (LIGHT_MIN, LIGHT_MAX),
    "moisture": (MOISTURE_MIN, MOISTURE_MAX),
    "temp": (TEMP_MIN, TEMP_MAX),
}

# === standard PSO parameters ===
PSO_W = 0.7
PSO_C1 = 1.5
PSO_C2 = 1.5
PSO_VMAX_RATIO = 0.2


def normalize_state(state):
    """Normalize the state to the [0, 1] range."""
    return np.array([
        (state[0] - BOUNDS["light"][0]) / (BOUNDS["light"][1] - BOUNDS["light"][0]),
        (state[1] - BOUNDS["moisture"][0]) / (BOUNDS["moisture"][1] - BOUNDS["moisture"][0]),
        (state[2] - BOUNDS["temp"][0]) / (BOUNDS["temp"][1] - BOUNDS["temp"][0]),
    ], dtype=np.float32)


def denormalize_state(normalized_state):
    """Restore the normalized state to the original range."""
    return np.array([
        normalized_state[0] * (BOUNDS["light"][1] - BOUNDS["light"][0]) + BOUNDS["light"][0],
        normalized_state[1] * (BOUNDS["moisture"][1] - BOUNDS["moisture"][0]) + BOUNDS["moisture"][0],
        normalized_state[2] * (BOUNDS["temp"][1] - BOUNDS["temp"][0]) + BOUNDS["temp"][0],
    ], dtype=float)


def generate_random_state():
    """Generate a random state within each dimension range."""
    light = np.random.randint(BOUNDS["light"][0], BOUNDS["light"][1] + 1)
    moisture = np.random.randint(BOUNDS["moisture"][0], BOUNDS["moisture"][1] + 1)
    temp = np.random.randint(BOUNDS["temp"][0], BOUNDS["temp"][1] + 1)
    return np.array([light, moisture, temp], dtype=float)


def clip_state(state):
    """Clip the state to each dimension boundary."""
    return np.array([
        np.clip(state[0], BOUNDS["light"][0], BOUNDS["light"][1]),
        np.clip(state[1], BOUNDS["moisture"][0], BOUNDS["moisture"][1]),
        np.clip(state[2], BOUNDS["temp"][0], BOUNDS["temp"][1]),
    ], dtype=float)


def is_state_valid(state):
    """Check whether the state is within all valid dimension ranges."""
    return (
        BOUNDS["light"][0] <= state[0] <= BOUNDS["light"][1]
        and BOUNDS["moisture"][0] <= state[1] <= BOUNDS["moisture"][1]
        and BOUNDS["temp"][0] <= state[2] <= BOUNDS["temp"][1]
    )


def execute_Tr(x, y, z):
    # 确保数值在合理范围内（外部范围已限定，但保留作为安全防护）
    x = max(STATE_MIN_X, min(STATE_MAX_X, x))
    y = max(STATE_MIN_Y, min(STATE_MAX_Y, y))
    z = max(STATE_MIN_Z, min(STATE_MAX_Z, z))

    triggered = set()

    if [(30 < x < 70)] != [(30 < x * 200 < 70)]: triggered.add(1)
    if [(30 < x < 70)] != [(30 < 6000 < 70)]: triggered.add(2)
    if [(30 < y < 70)] != [(30 < 7000 < 70)]: triggered.add(3)
    if [(30 < y < 70)] != [(30 < y * 600 < 70)]: triggered.add(4)
    if [(10 < z < 50)] != [(10 < 9000 < 50)]: triggered.add(5)
    if [(10 < z < 50)] != [(10 < z * 66 < 50)]: triggered.add(6)
    if [(x > 40 and x < 60 and y > 30 and y < 70)] != [(x > 40 and x < 60 and y > 30 and 9 < 70)]: triggered.add(7)
    if [(x > 35 and x < 65 and z > 20 and z < 40)] != [(x > 35 and x < 565 and z > 20 and z < 40)]: triggered.add(8)
    if [(x > 35 and x < 65 and z > 20 and z < 40)] != [(x > 35 and x * 11 < 65 and z > 20 and z < 40)]: triggered.add(9)
    if [(y > 30 and y < 70 and z > 20 and z < 40)] != [(y > 30 and y * 78 < 70 and z > 20 and z < 40)]: triggered.add(10)
    if [(y > 30 and y < 70 and z > 20 and z < 40)] != [(y > 30 and y < 70 and z * 56 > 20 and z < 40)]: triggered.add(11)
    if [(x > 45 and x < 55)] != [(x > 45 and x < 155)]: triggered.add(12)
    if [(y > 45 and y < 55)] != [(y > 45 and y < 88)]: triggered.add(13)
    if [(z > 25 and z < 35)] != [(z > 25 and z < 88)]: triggered.add(14)
    if [(z > 25 and z < 35)] != [(z > 288 and z < 35)]: triggered.add(15)
    if [(abs(x - 50) < 5)] != [(abs(x - 50) < 55)]: triggered.add(16)
    if [(abs(x - 50) < 5)] != [(abs(x - 50) < 533)]: triggered.add(17)
    if [(abs(y - 50) < 5)] != [(abs(y - 50) < 522)]: triggered.add(18)
    if [(abs(y - 50) < 5)] != [(abs(y - 50) < 95)]: triggered.add(19)
    if [(abs(z - 30) < 3)] != [(abs(z - 30) < 89)]: triggered.add(20)
    if [(abs(z - 30) < 3)] != [(abs(z - 30) < 63)]: triggered.add(21)
    if [(x > 30 + 5)] != [(x * 89 > 30 + 5)]: triggered.add(22)
    if [(x > 30 + 5)] != [(900 > 30 + 5)]: triggered.add(23)
    if [(y > 30 + 5)] != [(789 > 30 + 5)]: triggered.add(24)
    if [(y > 30 + 5)] != [(y * 78 > 30 + 5)]: triggered.add(25)
    if [(z > 10 + 5)] != [(999 > 10 + 5)]: triggered.add(26)
    if [(z > 10 + 5)] != [(z * 9 > 10 + 5)]: triggered.add(27)
    if [(x < 70 - 5)] != [(x * 88 < 70 - 5)]: triggered.add(28)
    if [(x < 70 - 5)] != [(988 < 70 - 5)]: triggered.add(29)
    if [(y < 70 - 5)] != [(y * 789 < 70 - 5)]: triggered.add(30)
    if [(y < 70 - 5)] != [(888 < 70 - 5)]: triggered.add(31)
    if [(z < 50 - 5)] != [(z * 96 < 50 - 5)]: triggered.add(32)
    if [(z < 50 - 5)] != [(777 < 50 - 5)]: triggered.add(33)
    if [(x > 35 and x < 65 and y > 35 and y < 65 and z > 15 and z < 45)] != [
        (x > 35 and x < 651 and y > 35 and y < 65 and z > 15 and z < 45)]: triggered.add(34)
    if [((x / 2 + y / 2 + z / 2) > 80 and (x / 2 + y / 2 + z / 2) < 120)] != [
        ((x / 2 + y / 2 + z / 2) > 810 and (x / 2 + y / 2 + z / 2) < 120)]: triggered.add(35)
    if [((x / 2 + y / 2 + z / 2) > 80 and (x / 2 + y / 2 + z / 2) < 120)] != [
        ((x / 2 + y / 2 + z / 2) > 1000 and (x / 2 + y / 2 + z / 2) < 120)]: triggered.add(36)
    if [(x / y > 0.8 and x / y < 1.2)] != [(x / y > 0.8 and x / y < 12)]: triggered.add(37)
    if [(x / z > 1.0 and x / z < 2.0)] != [(x / z > 1.0 and x / z < 20)]: triggered.add(38)
    if [(y / z > 1.0 and y / z < 2.0)] != [(y / z > 18 and y / z < 2.0)]: triggered.add(39)
    if [(y / z > 1.0 and y / z < 2.0)] != [(y / z > 1.0 and y / z < 20)]: triggered.add(40)
    if [(x * y > 1500 and x * y < 3500)] != [(x * y > 1500 and x * y < 350)]: triggered.add(41)
    if [(x * y > 1500 and x * y < 3500)] != [(x * y > 1500 and x * y < 35)]: triggered.add(42)
    if [(x * z > 500 and x * z < 2000)] != [(x * z > 5100 and x * z < 2000)]: triggered.add(43)
    if [(x * z > 500 and x * z < 2000)] != [(x * z > 5100 and x * z < 2000)]: triggered.add(44)
    if [(y * z > 500 and y * z < 2000)] != [(y * z > 5001 and y * z < 2000)]: triggered.add(45)
    if [(y * z > 500 and y * z < 2000)] != [(y * z > 500 and y * z < 200)]: triggered.add(46)
    if [((x + y) / 2 > 40 and (x + y) / 2 < 60)] != [((x + y) / 2 > 140 and (x + y) / 2 < 60)]: triggered.add(47)
    if [((x + y) / 2 > 40 and (x + y) / 2 < 60)] != [((x + y) / 2 > 40 and (x + y) / 2 < 6110)]: triggered.add(48)
    if [(x - y > -10 and x - y < 10)] != [(x - y > -10 and x - y < 110)]: triggered.add(49)
    if [(x - y > -10 and x - y < 10)] != [(x - y > -10 and x - y < 1220)]: triggered.add(50)
    if [(abs(x / y - 1.0) < 0.2)] != [(abs(x / y - 1.0) < 12)]: triggered.add(51)
    if [(abs(x / y - 1.0) < 0.2)] != [(abs(x / y - 1.0) < 211)]: triggered.add(52)
    if [(x / (y + 10) > 0.7 and x / (y + 10) < 1.3)] != [(x / (y + 10) > 711 and x / (y + 10) < 1.3)]: triggered.add(53)
    if [(x / (y + 10) > 0.7 and x / (y + 10) < 1.3)] != [(x / (y + 10) > 0.7 and x / (y + 10) < 13)]: triggered.add(54)
    if [(y / (z + 5) > 0.8 and y / (z + 5) < 1.5)] != [(y / (z + 5) > 0.8 and y / (z + 5) < 1225)]: triggered.add(55)
    if [(y / (z + 5) > 0.8 and y / (z + 5) < 1.5)] != [(y / (z + 5) > 0.8 and y / (z + 5) < 15)]: triggered.add(56)
    if [(z / (x + 5) > 0.5 and z / (x + 5) < 1.2)] != [(z / (x + 5) > 5 and z / (x + 5) < 1.2)]: triggered.add(57)
    if [(z / (x + 5) > 0.5 and z / (x + 5) < 1.2)] != [(z / (x + 5) > 511 and z / (x + 5) < 1.2)]: triggered.add(58)
    if [((x + y + z) / 3 > 30 and (x + y + z) / 3 < 50)] != [
        ((x + y + z) / 3 > 130 and (x + y + z) / 3 < 50)]: triggered.add(59)
    if [(math.sqrt((x - 50) ** 2 + (y - 50) ** 2) < 10)] != [
        (math.sqrt((x - 50) ** 2 + (y - 50) ** 2) < 1110)]: triggered.add(60)
    if [((x - 45) * (y - 45) > 100 and (x - 45) * (y - 45) < 1000)] != [
        ((x - 45) * (y - 45) > 100 and (x - 45) * (y - 45) < 200)]: triggered.add(61)
    if [((x - 45) * (y - 45) > 100 and (x - 45) * (y - 45) < 1000)] != [
        ((x - 45) * (y - 45) > 100 and (x - 45) * (y - 45) < 10)]: triggered.add(62)
    if [(abs((x + y) - 100) < 20)] != [(abs((x + y) - 100) < 201)]: triggered.add(63)
    if [(abs((x + y) - 100) < 20)] != [(abs((x + y) - 100) < 2110)]: triggered.add(64)
    if [(x / 50 > 0.8 and x / 50 < 1.2)] != [(x / 50 > 118 and x / 50 < 1.2)]: triggered.add(65)
    if [(x < 30)] != [(x < 220)]: triggered.add(66)
    if [(x > 70)] != [(x > 30)]: triggered.add(67)
    if [(x > 70)] != [(x > 440)]: triggered.add(68)
    if [(y < 30)] != [(y < 302)]: triggered.add(69)
    if [(y < 30)] != [(y < 130)]: triggered.add(70)
    if [(y > 70)] != [(y > 170)]: triggered.add(71)
    if [(y > 70)] != [(y > 7)]: triggered.add(72)
    if [(z < 15)] != [(z < 30)]: triggered.add(73)
    if [(z > 45)] != [(z > 450)]: triggered.add(74)
    if [(z > 45)] != [(z > 645)]: triggered.add(75)
    if [(x < 25 or x > 75)] != [(x < 25 or x > 575)]: triggered.add(76)
    if [(x < 25 or x > 75)] != [(x < 25 or x > 715)]: triggered.add(77)
    if [(y < 25 or y > 75)] != [(y < 25 or y > 751)]: triggered.add(78)
    if [(x > 65 and y > 65)] != [(x > 65 and y > 965)]: triggered.add(79)
    if [(x > 65 and y > 65)] != [(x > 65 and y > 635)]: triggered.add(80)
    if [(x < 35 and z < 20)] != [(x < 353 and z < 20)]: triggered.add(81)
    if [(x > 65 and z > 40)] != [(x > 65 and z > 140)]: triggered.add(82)
    if [(x > 65 and z > 40)] != [(x > 65 and z > 410)]: triggered.add(83)
    if [(y > 65 and z > 40)] != [(y > 651 and z > 40)]: triggered.add(84)
    if [(y > 65 and z > 40)] != [(y > 65 and z > 140)]: triggered.add(85)
    if [(abs(x / y - 1.0) > 0.4)] != [(abs(x / y - 1.0) > 4)]: triggered.add(86)
    if [(abs(x / y - 1.0) > 0.4)] != [(abs(x / y - 1.0) > 14)]: triggered.add(87)
    if [((x / 2 + y / 2 + z / 2) < 70 or (x / 2 + y / 2 + z / 2) > 130)] != [
        ((x / 2 + y / 2 + z / 2) < 7333 or (x / 2 + y / 2 + z / 2) > 130)]: triggered.add(88)
    if [((x / 2 + y / 2 + z / 2) < 70 or (x / 2 + y / 2 + z / 2) > 130)] != [
        ((x / 2 + y / 2 + z / 2) < 788 or (x / 2 + y / 2 + z / 2) > 130)]: triggered.add(89)
    if [(x < 20 or x > 80)] != [(x < 120 or x > 80)]: triggered.add(90)
    if [(x < 20 or x > 80)] != [(x < 20 or x > 99)]: triggered.add(91)
    if [(z > 27 and z < 33)] != [(z > 27 and z < 133)]: triggered.add(92)
    if [(z > 27 and z < 33)] != [(z > 27 and z < 331)]: triggered.add(93)
    if [(abs(z - 30) < 1)] != [(abs(z - 30) < 61)]: triggered.add(94)
    if [(abs(z - 30) < 1)] != [(abs(z - 30) < 133)]: triggered.add(95)
    if [(x / y > 0.95 and x / y < 1.05)] != [(x / y > 0.95 and x / y < 105)]: triggered.add(96)
    if [((x / 2 + y / 2 + z / 2) > 90 and (x / 2 + y / 2 + z / 2) < 110)] != [
        ((x / 2 + y / 2 + z / 2) > 999 and (x / 2 + y / 2 + z / 2) < 110)]: triggered.add(97)
    if [(abs((x + y) / 2 - 50) < 2)] != [(abs((x + y) / 2 - 50) < 112)]: triggered.add(98)
    if [(abs((x + y) / 2 - 50) < 2)] != [(abs((x + y) / 2 - 50) < 222)]: triggered.add(99)
    if [(x / z > 1.2 and x / z < 1.8)] != [(x / z > 1.2 and x / z < 18)]: triggered.add(100)
    if [(x / z > 1.2 and x / z < 1.8)] != [(x / z > 1.2 and x / z < 188)]: triggered.add(101)
    if [(y / z > 1.2 and y / z < 1.8)] != [(y / z > 12 and y / z < 1.8)]: triggered.add(102)
    if [(y / z > 1.2 and y / z < 1.8)] != [(y / z > 122 and y / z < 1.8)]: triggered.add(103)
    if [((y - 45) / 10 > 0.45 and (y - 45) / 10 < 0.55)] != [
        ((y - 45) / 10 > 0.45 and (y - 45) / 10 < 755)]: triggered.add(104)
    if [((y - 45) / 10 > 0.45 and (y - 45) / 10 < 0.55)] != [
        ((y - 45) / 10 > 0.45 and (y - 45) / 10 < 55)]: triggered.add(105)
    if [(min(z - 15, 45 - z) > 5)] != [(min(z - 15, 45 - z) > 511)]: triggered.add(106)
    if [(min(z - 15, 45 - z) > 5)] != [(min(z - 15, 45 - z) > 5220)]: triggered.add(107)
    if [(abs(x * y - 2500) < 50)] != [(abs(x * y - 2500) < 5012)]: triggered.add(108)

    return triggered


target_paths = [
    {3, 4, 5, 6, 8, 10, 12, 13, 14, 16, 17, 18, 19, 20, 21, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 45, 46, 48, 49, 50,
     51, 52, 54, 60, 61, 62, 63, 64, 66, 68, 69, 70, 72, 76, 77, 86, 87, 88, 89, 91, 92, 93, 94, 95, 96, 97, 98, 99,
     100, 101, 102, 103, 104, 105, 106, 107, 108},

    {3, 4, 5, 6, 8, 10, 12, 13, 14, 16, 17, 18, 19, 20, 21, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 48, 49, 50, 51, 52,
     53, 55, 56, 60, 61, 62, 63, 64, 66, 68, 69, 70, 72, 76, 77, 86, 87, 88, 89, 91, 92, 93, 94, 95, 96, 97, 98, 99,
     100, 101, 102, 103, 104, 105, 106, 107, 108},

    {3, 4, 5, 6, 8, 10, 12, 13, 15, 16, 17, 18, 19, 20, 21, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 45, 46, 48, 49, 50,
     51, 52, 53, 60, 61, 62, 63, 64, 66, 68, 69, 70, 72, 76, 77, 86, 87, 88, 89, 90, 92, 93, 94, 95, 96, 98, 99, 100,
     101, 102, 103, 104, 105, 106, 107, 108},

    {3, 4, 5, 6, 8, 10, 12, 13, 15, 16, 17, 18, 19, 20, 21, 30, 31, 32, 33, 34, 35, 36, 37, 38, 40, 45, 46, 48, 49, 50,
     51, 52, 54, 55, 56, 60, 61, 62, 63, 64, 66, 68, 69, 70, 72, 73, 76, 77, 86, 87, 88, 89, 91, 94, 95, 96, 98, 99,
     100, 101, 104, 105, 106, 107, 108},

    {3, 4, 5, 6, 11, 12, 13, 16, 17, 18, 19, 20, 21, 26, 27, 30, 31, 32, 33, 35, 36, 37, 38, 40, 43, 44, 45, 46, 48, 49,
     50, 51, 52, 54, 55, 56, 60, 61, 62, 63, 64, 66, 68, 69, 70, 72, 73, 76, 77, 81, 86, 87, 88, 89, 91, 94, 95, 96, 98,
     99, 100, 101, 104, 105, 108},

    {3, 4, 5, 6, 8, 10, 12, 13, 14, 16, 17, 18, 19, 20, 21, 32, 33, 35, 36, 37, 38, 39, 48, 49, 50, 51, 52, 53, 55, 56,
     60, 61, 62, 63, 64, 66, 68, 69, 70, 72, 76, 77, 79, 80, 88, 89, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100, 101, 102,
     103, 104, 105, 106, 107, 108},

    {3, 4, 5, 6, 12, 13, 14, 16, 17, 18, 19, 20, 21, 32, 33, 35, 36, 37, 38, 39, 48, 49, 50, 51, 52, 53, 55, 56, 60, 61,
     62, 63, 64, 66, 68, 69, 70, 72, 76, 77, 79, 80, 82, 83, 84, 85, 88, 89, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100,
     101, 102, 103, 104, 105, 108},

    {3, 4, 5, 6, 12, 13, 14, 16, 17, 18, 19, 20, 21, 30, 31, 32, 33, 34, 35, 36, 37, 39, 48, 49, 50, 51, 52, 53, 57, 58,
     60, 61, 62, 63, 64, 66, 68, 69, 70, 72, 76, 77, 82, 83, 86, 87, 88, 89, 90, 92, 93, 94, 95, 96, 98, 99, 100, 101,
     102, 103, 104, 105, 108},

    {3, 4, 5, 6, 12, 13, 14, 16, 17, 18, 19, 20, 21, 35, 36, 37, 39, 48, 49, 50, 51, 52, 53, 57, 58, 60, 61, 62, 63, 64,
     66, 68, 69, 70, 72, 74, 75, 76, 77, 79, 80, 82, 83, 84, 85, 88, 89, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100, 101,
     102, 103, 104, 105, 108},

    {3, 4, 5, 6, 8, 10, 12, 16, 17, 18, 19, 20, 21, 24, 25, 30, 31, 32, 33, 37, 38, 39, 41, 42, 43, 44, 45, 46, 48, 49,
     50, 51, 52, 54, 59, 60, 63, 64, 66, 68, 69, 70, 72, 73, 76, 77, 86, 87, 88, 89, 91, 94, 95, 96, 98, 99, 100, 101,
     102, 103, 106, 107, 108},

    {5, 6, 12, 13, 14, 16, 17, 18, 19, 20, 21, 32, 33, 35, 36, 37, 38, 39, 48, 49, 50, 51, 52, 53, 55, 56, 60, 63, 64,
     66, 68, 69, 70, 71, 76, 77, 78, 79, 80, 82, 83, 84, 85, 88, 89, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100, 101, 102,
     103, 104, 105, 108},

    {1, 2, 3, 4, 5, 6, 9, 10, 12, 16, 17, 18, 19, 20, 21, 24, 25, 28, 29, 30, 31, 32, 33, 37, 38, 39, 41, 42, 43, 44,
     45, 46, 47, 49, 50, 51, 52, 53, 59, 60, 65, 66, 67, 69, 70, 72, 73, 86, 87, 90, 94, 95, 96, 98, 99, 100, 101, 102,
     103, 106, 107, 108},

    {1, 2, 5, 6, 7, 9, 12, 15, 16, 17, 18, 19, 20, 21, 28, 29, 32, 33, 35, 36, 38, 40, 43, 44, 48, 51, 52, 55, 56, 60,
     61, 62, 63, 64, 65, 66, 67, 69, 70, 71, 73, 78, 86, 87, 88, 89, 90, 94, 95, 98, 99, 100, 101, 104, 105, 106, 107,
     108},

    {1, 2, 3, 4, 5, 6, 10, 13, 14, 16, 17, 18, 19, 20, 21, 22, 23, 28, 29, 30, 31, 32, 33, 39, 41, 42, 43, 44, 47, 51,
     52, 55, 56, 57, 58, 59, 60, 66, 67, 69, 70, 72, 86, 87, 90, 92, 93, 94, 95, 98, 99, 102, 103, 104, 105, 106, 107,
     108},
]


def jaccard_similarity(set1: Set[int], set2: Set[int]) -> float:
    """If set1 covers target set2, the match score is treated as 1."""
    if set2.issubset(set1):
        return 1.0
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    return intersection / union if union != 0 else 0.0


def compute_reward(state, target_path, triggered):
    """Compute the DQN reward."""
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward


class SimpleReplayBuffer:
    def __init__(self, capacity=5000):
        self.buffer = deque(maxlen=capacity)

    def append(self, experience):
        self.buffer.append(experience)

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return random.sample(self.buffer, len(self.buffer))
        return random.sample(self.buffer, batch_size)

    def __len__(self):
        return len(self.buffer)


class SimpleDQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super().__init__()
        self.fc1 = nn.Linear(state_dim, 64)
        self.fc2 = nn.Linear(64, 32)
        self.fc3 = nn.Linear(32, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)


class SimpleDQNAgent:
    def __init__(self, state_dim, action_dim, learning_rate=0.001):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.epsilon = 0.3
        self.gamma = 0.99

        self.model = SimpleDQN(state_dim, action_dim).to(device)
        self.target_model = SimpleDQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())
        self.replay_buffer = SimpleReplayBuffer(capacity=5000)

    def decode_action(self, action_idx):
        """Decode the action."""
        light_deltas = [int((LIGHT_MAX - LIGHT_MIN) * p) for p in [0.05, 0.1, 0.2, 0.5, 0.7, -0.05, -0.1, -0.2, -0.5, -0.7]]
        moisture_deltas = [int((MOISTURE_MAX - MOISTURE_MIN) * p) for p in [0.05, 0.1, 0.2, 0.5, 0.7, -0.05, -0.1, -0.2, -0.5, -0.7]]
        temp_deltas = [int((TEMP_MAX - TEMP_MIN) * p) for p in [0.05, 0.1, 0.2, 0.5, 0.7, -0.05, -0.1, -0.2, -0.5, -0.7]]

        dim = action_idx // 10
        delta_idx = action_idx % 10

        if dim == 0:
            return (light_deltas[delta_idx], 0, 0)
        if dim == 1:
            return (0, moisture_deltas[delta_idx], 0)
        return (0, 0, temp_deltas[delta_idx])

    def store_transition(self, state, action, reward, next_state, done):
        self.replay_buffer.append((state, action, reward, next_state, done))

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return

        batch = self.replay_buffer.sample(batch_size)
        states, actions, rewards, next_states, dones = zip(*batch)

        states = torch.tensor(np.array(states), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array(next_states), dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))

        loss = nn.MSELoss()(current_q_values, target_q_values)

        self.optimizer.zero_grad()
        loss.backward()
        self.optimizer.step()

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())

    def get_best_samples(self, target_path, num_samples=20):
        """Screen DQN candidate samples from the replay buffer. This stage is counted separately in T_sample_screening."""
        if len(self.replay_buffer) == 0:
            return []

        samples_with_scores = []
        for state_norm, _, _, _, _ in self.replay_buffer.buffer:
            state_original = denormalize_state(state_norm)
            triggered = execute_Tr(state_original)
            reward = compute_reward(state_original, target_path, triggered)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_scores.append((state_original, reward, sim, triggered))

        samples_with_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_scores[:num_samples]


def train_dqn_for_path(path_idx, target_path, num_samples=200):
    """Train a standard DQN for a single path."""
    print(f"  Start training path {path_idx + 1} DQN model...")
    agent = SimpleDQNAgent(state_dim=3, action_dim=30)

    random_states = [normalize_state(generate_random_state()) for _ in range(num_samples)]

    STEPS_PER_SAMPLE = 3
    EPOCHS = 5
    BATCH_SIZE = 32

    step_count = 0
    for _ in range(EPOCHS):
        for state_norm in random_states:
            state_norm = tuple(state_norm)
            state_original = denormalize_state(state_norm)

            for step in range(STEPS_PER_SAMPLE):
                legal_actions = []
                for a in range(agent.action_dim):
                    dx, dy, dz = agent.decode_action(a)
                    next_state_candidate = (state_original[0] + dx, state_original[1] + dy, state_original[2] + dz)
                    if is_state_valid(next_state_candidate):
                        legal_actions.append(a)

                if not legal_actions:
                    break

                if random.random() < agent.epsilon:
                    action = random.choice(legal_actions)
                else:
                    state_tensor = torch.tensor(state_norm, dtype=torch.float32).unsqueeze(0).to(device)
                    with torch.no_grad():
                        q_values = agent.model(state_tensor)[0]
                    action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                dx, dy, dz = agent.decode_action(action)
                next_state_original = (state_original[0] + dx, state_original[1] + dy, state_original[2] + dz)
                next_state_norm = normalize_state(next_state_original)

                triggered = execute_Tr(next_state_original)
                reward = compute_reward(next_state_original, target_path, triggered)
                done = (step == STEPS_PER_SAMPLE - 1)

                agent.store_transition(state_norm, action, reward, next_state_norm, done)

                state_norm = next_state_norm
                state_original = next_state_original
                step_count += 1

                if step_count % 50 == 0 and len(agent.replay_buffer) >= BATCH_SIZE:
                    agent.train(BATCH_SIZE)

                if step_count % 100 == 0:
                    agent.update_target_model()

    print(f"  Path {path_idx + 1} training completed, replay buffer size: {len(agent.replay_buffer)}")
    return agent


class Particle:
    def __init__(self, initial_position=None):
        if initial_position is not None:
            self.position = clip_state(np.array(initial_position, dtype=float))
        else:
            self.position = np.array([
                np.random.uniform(BOUNDS["light"][0], BOUNDS["light"][1]),
                np.random.uniform(BOUNDS["moisture"][0], BOUNDS["moisture"][1]),
                np.random.uniform(BOUNDS["temp"][0], BOUNDS["temp"][1]),
            ], dtype=float)

        spans = np.array([
            BOUNDS["light"][1] - BOUNDS["light"][0],
            BOUNDS["moisture"][1] - BOUNDS["moisture"][0],
            BOUNDS["temp"][1] - BOUNDS["temp"][0],
        ], dtype=float)
        self.velocity = np.random.uniform(-0.1 * spans, 0.1 * spans)
        self.best_position = self.position.copy()
        self.best_fitness = -1.0
        self.fitness = -1.0


class BasicPSO:
    """
    Standard PSO: contains only standard velocity and position updates.
    Mutation coefficients, opposite particles, local-best reset, and other improved PSO mechanisms have been removed.
    """

    def __init__(self, target_path, swarm_size=20, dqn_samples=None):
        self.target_path = target_path
        self.swarm_size = swarm_size
        self.particles = []
        self.global_best_position = None
        self.global_best_fitness = -1.0

        if dqn_samples is not None and len(dqn_samples) > 0:
            num_direct = min(len(dqn_samples), swarm_size)
            for i in range(num_direct):
                state_tuple, _, _, _ = dqn_samples[i]
                self.particles.append(Particle(initial_position=state_tuple))

            while len(self.particles) < swarm_size:
                base_idx = len(self.particles) % len(dqn_samples)
                state_tuple, _, _, _ = dqn_samples[base_idx]
                perturbed = np.array(state_tuple, dtype=float) + np.random.randint(-10, 11, size=3)
                self.particles.append(Particle(initial_position=clip_state(perturbed)))
        else:
            self.particles = [Particle() for _ in range(swarm_size)]

        for particle in self.particles:
            particle.fitness = self.fitness_function(particle.position)
            particle.best_fitness = particle.fitness
            particle.best_position = particle.position.copy()

            if particle.fitness > self.global_best_fitness:
                self.global_best_fitness = particle.fitness
                self.global_best_position = particle.position.copy()

    def fitness_function(self, position):
        try:
            triggered = execute_Tr(position)
            return jaccard_similarity(triggered, self.target_path)
        except Exception:
            return 0.0

    def update(self):
        spans = np.array([
            BOUNDS["light"][1] - BOUNDS["light"][0],
            BOUNDS["moisture"][1] - BOUNDS["moisture"][0],
            BOUNDS["temp"][1] - BOUNDS["temp"][0],
        ], dtype=float)
        max_velocity = PSO_VMAX_RATIO * spans

        for particle in self.particles:
            r1 = np.random.random(3)
            r2 = np.random.random(3)

            particle.velocity = (
                PSO_W * particle.velocity
                + PSO_C1 * r1 * (particle.best_position - particle.position)
                + PSO_C2 * r2 * (self.global_best_position - particle.position)
            )
            particle.velocity = np.clip(particle.velocity, -max_velocity, max_velocity)

            particle.position = clip_state(particle.position + particle.velocity)
            particle.fitness = self.fitness_function(particle.position)

            if particle.fitness > particle.best_fitness:
                particle.best_fitness = particle.fitness
                particle.best_position = particle.position.copy()

            if particle.fitness > self.global_best_fitness:
                self.global_best_fitness = particle.fitness
                self.global_best_position = particle.position.copy()


def _set_header(ws, headers, widths, header_fill, header_font, border, center_align):
    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_row(ws, row_idx, row_data, border, alignment, fill=None):
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.border = border
        cell.alignment = alignment
        if fill is not None:
            cell.fill = fill


def export_time_metrics_to_excel(all_run_results, experiment_total_time=None, filename=None):
    """Export only timing metrics and per-path iteration counts."""
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"DQN_BasicPSO_Time_Metrics_{timestamp}.xlsx"

    output_dir = os.path.join(os.getcwd(), "results")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    num_runs = len(all_run_results)
    num_paths = len(target_paths)

    # Worksheet 1: runtime summary
    ws1 = wb.active
    ws1.title = "Runtime Summary"
    ws1.sheet_view.showGridLines = False
    headers1 = [
        "Run", "Number of Paths", "DQNTraining Total Time(s)", "Sample Screening Total Time(s)", "DQNDirect Check Total Time(s)",
        "PSOInitialization Total Time(s)", "PSOSearch Total Time(s)", "Total Algorithm Time(s)", "Total Iterations", "Average Iterations per Path"
    ]
    widths1 = [12, 10, 18, 18, 22, 18, 18, 16, 14, 18]
    _set_header(ws1, headers1, widths1, header_fill, header_font, border, center_align)

    for run_idx, run_result in enumerate(all_run_results, 1):
        path_results = run_result["paths"]
        total_dqn_train = sum(p["T_DQN_train"] for p in path_results)
        total_screen = sum(p["T_sample_screening"] for p in path_results)
        total_direct_check = sum(p["T_direct_check"] for p in path_results)
        total_pso_init = sum(p["T_PSO_init"] for p in path_results)
        total_pso_search = sum(p["T_PSO_search"] for p in path_results)
        total_iterations = sum(p["iterations"] for p in path_results)

        row_data = [
            f" {run_idx}",
            num_paths,
            round(total_dqn_train, 6),
            round(total_screen, 6),
            round(total_direct_check, 6),
            round(total_pso_init, 6),
            round(total_pso_search, 6),
            round(run_result["T_run_algorithm_total"], 6),
            int(total_iterations),
            round(total_iterations / num_paths, 2),
        ]
        _write_row(ws1, run_idx + 1, row_data, border, center_align, alternate_fill if run_idx % 2 == 0 else None)

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:J{num_runs + 1}"

    # Worksheet 2: path timing and iteration details
    ws2 = wb.create_sheet(title="Path Timing and Iteration Details")
    ws2.sheet_view.showGridLines = False
    headers2 = [
        "Run", "Path ID", "DQN training(s)", "Sample Screening Time(s)", "DQNDirect Check Time(s)",
        "PSOInitialization Time(s)", "PSOSearch Time(s)", "Path Total Algorithm Time(s)", "Iterations"
    ]
    widths2 = [12, 12, 18, 18, 22, 18, 18, 18, 12]
    _set_header(ws2, headers2, widths2, header_fill, header_font, border, center_align)

    row_idx = 2
    for run_idx, run_result in enumerate(all_run_results, 1):
        for path_result in run_result["paths"]:
            row_data = [
                f" {run_idx}",
                f"Path  {path_result['path_idx'] + 1}",
                round(path_result["T_DQN_train"], 6),
                round(path_result["T_sample_screening"], 6),
                round(path_result["T_direct_check"], 6),
                round(path_result["T_PSO_init"], 6),
                round(path_result["T_PSO_search"], 6),
                round(path_result["T_path_algorithm_total"], 6),
                int(path_result["iterations"]),
            ]
            _write_row(ws2, row_idx, row_data, border, center_align, alternate_fill if row_idx % 2 == 0 else None)
            row_idx += 1

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:I{row_idx - 1}"

    # Worksheet 3: path aggregate statistics
    ws3 = wb.create_sheet(title="Path Aggregate Statistics")
    ws3.sheet_view.showGridLines = False
    headers3 = [
        "Path ID", "Average DQN Training Time(s)", "Average Sample Screening Time(s)", "Average PSO Initialization Time(s)",
        "Average PSO Search Time(s)", "Average Path Total Algorithm Time(s)", "Average Iterations", "Minimum Iterations", "Maximum Iterations"
    ]
    widths3 = [12, 22, 22, 22, 22, 22, 16, 14, 14]
    _set_header(ws3, headers3, widths3, header_fill, header_font, border, center_align)

    for path_idx in range(num_paths):
        records = [run_result["paths"][path_idx] for run_result in all_run_results]
        iterations = [r["iterations"] for r in records]
        row_data = [
            f"Path  {path_idx + 1}",
            round(float(np.mean([r["T_DQN_train"] for r in records])), 6),
            round(float(np.mean([r["T_sample_screening"] for r in records])), 6),
            round(float(np.mean([r["T_PSO_init"] for r in records])), 6),
            round(float(np.mean([r["T_PSO_search"] for r in records])), 6),
            round(float(np.mean([r["T_path_algorithm_total"] for r in records])), 6),
            round(float(np.mean(iterations)), 2),
            int(np.min(iterations)),
            int(np.max(iterations)),
        ]
        _write_row(ws3, path_idx + 2, row_data, border, center_align, alternate_fill if (path_idx + 1) % 2 == 0 else None)

    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:I{num_paths + 1}"

    # Worksheet 4: experiment overview
    ws4 = wb.create_sheet(title="Experiment Overview")
    ws4.sheet_view.showGridLines = False
    headers4 = ["Metric", "Value"]
    widths4 = [32, 24]
    _set_header(ws4, headers4, widths4, header_fill, header_font, border, center_align)

    all_path_records = [p for run_result in all_run_results for p in run_result["paths"]]
    all_iterations = [p["iterations"] for p in all_path_records]
    overview_rows = [
        ["Experiment Runs", num_runs],
        ["Number of Target Paths", num_paths],
        ["Total Experiment Wall-clock Time(s)", round(experiment_total_time, 6) if experiment_total_time is not None else ""],
        ["Average Total Algorithm Time per Run(s)", round(float(np.mean([r["T_run_algorithm_total"] for r in all_run_results])), 6)],
        ["Average Total Algorithm Time per Path(s)", round(float(np.mean([p["T_path_algorithm_total"] for p in all_path_records])), 6)],
        ["Average DQN Training Time per Path(s)", round(float(np.mean([p["T_DQN_train"] for p in all_path_records])), 6)],
        ["Average Sample Screening Time per Path(s)", round(float(np.mean([p["T_sample_screening"] for p in all_path_records])), 6)],
        ["Average PSO Search Time per Path(s)", round(float(np.mean([p["T_PSO_search"] for p in all_path_records])), 6)],
        ["Average Iterations per Path", round(float(np.mean(all_iterations)), 2)],
        ["Maximum Iterations per Path", int(np.max(all_iterations))],
        ["Minimum Iterations per Path", int(np.min(all_iterations))],
    ]

    for idx, row_data in enumerate(overview_rows, 2):
        _write_row(ws4, idx, row_data, border, center_align, alternate_fill if idx % 2 == 0 else None)

    wb.save(filepath)
    print(f"\n Timing metrics exported to: {filepath}")
    return filepath


def run_single_experiment(run_num, max_iterations=3000):
    """Run one DQN + standard PSO ."""
    print(f"\n{'=' * 80}")
    print(f"Start run  {run_num}  run: DQN + standard PSO ")
    print(f"{'=' * 80}")

    run_start = time.perf_counter()
    path_results = []

    for path_idx, target_path in enumerate(target_paths):
        print(f"\nRun {run_num} run - Path {path_idx + 1}")
        path_total_start = time.perf_counter()

        # 1) DQN training
        dqn_train_start = time.perf_counter()
        agent = train_dqn_for_path(path_idx, target_path, num_samples=200)
        T_DQN_train = time.perf_counter() - dqn_train_start

        # 2) DQNSample Screening Time
        screening_start = time.perf_counter()
        dqn_samples = agent.get_best_samples(target_path, num_samples=20)
        T_sample_screening = time.perf_counter() - screening_start

        # 3) DQN
        direct_check_start = time.perf_counter()
        direct_solution_found = False
        for state_tuple, reward, sim, triggered in dqn_samples:
            if sim >= 1.0 or target_path.issubset(triggered):
                direct_solution_found = True
                break
        T_direct_check = time.perf_counter() - direct_check_start

        T_PSO_init = 0.0
        T_PSO_search = 0.0
        iterations_used = 0

        if direct_solution_found:
            print(
                f"  Path {path_idx + 1}: DQN | "
                f"DQN training {T_DQN_train:.6f}s | screening {T_sample_screening:.6f}s | iterations {iterations_used}"
            )
        else:
            # 4) standard PSOInitialization Time
            pso_init_start = time.perf_counter()
            pso = BasicPSO(target_path, swarm_size=20, dqn_samples=dqn_samples)
            T_PSO_init = time.perf_counter() - pso_init_start

            # 5) standard PSOSearch Time
            pso_search_start = time.perf_counter()
            iterations_used = max_iterations
            for iteration in range(max_iterations):
                pso.update()
                if pso.global_best_fitness >= 1.0:
                    iterations_used = iteration + 1
                    break
            T_PSO_search = time.perf_counter() - pso_search_start

            print(
                f"  Path {path_idx + 1}: "
                f"DQN training {T_DQN_train:.6f}s | screening {T_sample_screening:.6f}s | "
                f"PSO {T_PSO_init:.6f}s | PSO {T_PSO_search:.6f}s | "
                f"iterations {iterations_used}"
            )

        T_path_algorithm_total = time.perf_counter() - path_total_start

        path_results.append({
            "path_idx": path_idx,
            "T_DQN_train": T_DQN_train,
            "T_sample_screening": T_sample_screening,
            "T_direct_check": T_direct_check,
            "T_PSO_init": T_PSO_init,
            "T_PSO_search": T_PSO_search,
            "T_path_algorithm_total": T_path_algorithm_total,
            "iterations": iterations_used,
        })

    T_run_algorithm_total = time.perf_counter() - run_start
    print(f"\nRun {run_num} runcompleted | Total Algorithm Time {T_run_algorithm_total:.6f}s")
    return {
        "run_num": run_num,
        "paths": path_results,
        "T_run_algorithm_total": T_run_algorithm_total,
    }


def run_multiple_experiments(num_runs):
    print("\n" + "=" * 80)
    print(f"DQN + standard PSO - {num_runs}: Metric")
    print("=" * 80)
    print(f"Number of Target Paths: {len(target_paths)}")
    print("Statistics: DQN training, Sample Screening Time, PSOInitialization Time, PSOSearch Time, Path Iterations")
    print("=" * 80)

    all_run_results = []
    experiment_start = time.perf_counter()

    for run_num in range(1, num_runs + 1):
        run_result = run_single_experiment(run_num)
        all_run_results.append(run_result)

    experiment_total_time = time.perf_counter() - experiment_start
    print(f"\n{'=' * 80}")
    print(f"All {num_runs} runcompleted | Total Experiment Wall-clock Time {experiment_total_time:.6f}s")
    print(f"{'=' * 80}\n")
    return all_run_results, experiment_total_time


if __name__ == "__main__":
    print("=" * 80)
    print("DQN + standard PSO - MetricPath Iterations")
    print("=" * 80)
    print(f"Current configuration: Run = {NUM_RUNS}")
    print(f"Number of Paths: {len(target_paths)}")
    print(f"Device: {device}")
    print("=" * 80)

    if len(sys.argv) > 1:
        try:
            NUM_RUNS = int(sys.argv[1])
            print(f"Read from command line: Run = {NUM_RUNS}")
        except ValueError:
            print(f"Invalid command-line argument, using default number of runs {NUM_RUNS}")

    all_results, total_time = run_multiple_experiments(num_runs=NUM_RUNS)
    export_time_metrics_to_excel(all_results, total_time)
    print("\nProgram completed")
