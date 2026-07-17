import os
import random
import math
import numpy as np
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# ================================================================
# 1. 状态范围配置（匹配 section8 的物理含义）
# ================================================================
STATE_MIN_X, STATE_MAX_X = 1000, 2000   # 速度
STATE_MIN_Y, STATE_MAX_Y = 100, 200     # 扭矩
STATE_MIN_Z, STATE_MAX_Z = 10, 30       # 电流

def generate_input():
    # 使用 uniform 生成连续值，因为 section8 的规则涉及浮点比较
    return [
        random.uniform(STATE_MIN_X, STATE_MAX_X),
        random.uniform(STATE_MIN_Y, STATE_MAX_Y),
        random.uniform(STATE_MIN_Z, STATE_MAX_Z)
    ]


def section8_hybrid_speed_torque_current(x, y, z):
    """第8类: 速度扭矩电流混合控制 (速度x, 扭矩y, 电流z) - 144个有效变异分支"""
    triggered = set()  # 记录触发的条件编号（1-144）


    # -------------------------- 1-32：速度扭矩电流基础检查 --------------------------
    # 编号1 (原编号7)
    if [(x > 1400 and x < 1600 and y > 160 and y < 190)] != [(x > 1400 and x * 8 < 1600 and y > 160 and y < 190)]:
        triggered.add(1)
    # 编号2 (原编号8)
    if [(x > 1400 and x < 1600 and y > 160 and y < 190)] != [(x > 1400 and x < 1600 and y * 10 > 160 and y < 190)]:
        triggered.add(2)
    # 编号3 (原编号9)
    if [(x > 1350 and x < 1650 and z > 11 and z < 14)] != [(x > 1350 and x * 9 < 1650 and z > 11 and z < 14)]:
        triggered.add(3)
    # 编号4 (原编号10)
    if [(x > 1350 and x < 1650 and z > 11 and z < 14)] != [(x > 1350 and x < 1650 and z * 20 > 11 and z < 14)]:
        triggered.add(4)
    # 编号5 (原编号11)
    if [(y > 155 and y < 195 and z > 11 and z < 14)] != [(y > 155 and y * 7 < 195 and z > 11 and z < 14)]:
        triggered.add(5)
    # 编号6 (原编号12)
    if [(y > 155 and y < 195 and z > 11 and z < 14)] != [(y > 155 and y < 195 and z * 22 > 11 and z < 14)]:
        triggered.add(6)
    # 编号7 (原编号13)
    if [(x > 1480 and x < 1520)] != [(x > 1480 and x * 8 < 1520)]:
        triggered.add(7)
    # 编号8 (原编号14)
    if [(x > 1480 and x < 1520)] != [(x * 12 > 1480 and x < 1520)]:
        triggered.add(8)
    # 编号9 (原编号15)
    if [(y > 173 and y < 177)] != [(y * 12 > 173 and y < 177)]:
        triggered.add(9)
    # 编号10 (原编号16)
    if [(y > 173 and y < 177)] != [(y > 173 and y * 22 < 177)]:
        triggered.add(10)
    # 编号11 (原编号17)
    if [(z > 12.2 and z < 12.8)] != [(z * 78 > 12.2 and z < 12.8)]:
        triggered.add(11)
    # 编号12 (原编号18)
    if [(z > 12.2 and z < 12.8)] != [(66 > 12.2 and z < 12.8)]:
        triggered.add(12)
    # 编号13 (原编号19)
    if [(abs(x - 1500) < 40)] != [(abs(x - 1500) < 140)]:
        triggered.add(13)
    # 编号14 (原编号20)
    if [(abs(x - 1500) < 40)] != [(abs(x - 1500) < 990)]:
        triggered.add(14)
    # 编号15 (原编号21)
    if [(abs(y - 175) < 6)] != [(abs(y - 175) < 16)]:
        triggered.add(15)
    # 编号16 (原编号22)
    if [(abs(y - 175) < 6)] != [(abs(y - 175) < 226)]:
        triggered.add(16)
    # 编号17 (原编号23)
    if [(abs(z - 12.5) < 0.4)] != [(abs(z - 12.5) < 4)]:
        triggered.add(17)
    # 编号18 (原编号24)
    if [(abs(z - 12.5) < 0.4)] != [(abs(z - 12.5) < 1)]:
        triggered.add(18)
    # 编号19 (原编号25)
    if [(x > 1350 and y > 155 and z > 11)] != [(x > 1350 and y * 78 > 155 and z > 11)]:
        triggered.add(19)
    # 编号20 (原编号26)
    if [(x > 1350 and y > 155 and z > 11)] != [(x * 67 > 1350 and y > 155 and z > 11)]:
        triggered.add(20)
    # 编号21 (原编号27)
    if [(x < 1650 and y < 195 and z < 14)] != [(x < 1650 and y * 88 < 195 and z < 14)]:
        triggered.add(21)
    # 编号22 (原编号28)
    if [(x < 1650 and y < 195 and z < 14)] != [(x < 1650 and y < 195 and z * 8 < 14)]:
        triggered.add(22)
    # 编号23 (原编号29)
    if [((x / 10 + y + z * 10) > 350 and (x / 10 + y + z * 10) < 410)] != [
        ((x / 10 + y + z * 10) > 3150 and (x / 10 + y + z * 10) < 410)]:
        triggered.add(23)
    # 编号24 (原编号30)
    if [((x / 10 + y + z * 10) > 350 and (x / 10 + y + z * 10) < 410)] != [
        ((x / 10 + y + z * 10) > 350 and (x / 10 + y + z * 10) < 1410)]:
        triggered.add(24)
    # 编号25 (原编号31)
    if [(x * y / 1000 > 230 and x * y / 1000 < 310)] != [(x * y / 1000 > 1230 and x * y / 1000 < 310)]:
        triggered.add(25)
    # 编号26 (原编号32)
    if [(x * y / 1000 > 230 and x * y / 1000 < 310)] != [(x * y / 1000 > 230 and x * y / 1000 < 3110)]:
        triggered.add(26)
    # 编号27 (原编号33)
    if [(x * z / 100 > 160 and x * z / 100 < 220)] != [(x * z / 100 > 1610 and x * z / 100 < 220)]:
        triggered.add(27)
    # 编号28 (原编号34)
    if [(x * z / 100 > 160 and x * z / 100 < 220)] != [(x * z / 100 > 160 and x * z / 100 < 2120)]:
        triggered.add(28)
    # 编号29 (原编号35)
    if [(y * z > 2000 and y * z < 2600)] != [(y * z > 200 and y * z < 2600)]:
        triggered.add(29)
    # 编号30 (原编号36)
    if [(y * z > 2000 and y * z < 2600)] != [(y * z > 2000 and y * z < 260)]:
        triggered.add(30)
    # 编号31 (原编号37)
    if [((x / 10 + y + z * 10) / 3 > 116 and (x / 10 + y + z * 10) / 3 < 136)] != [
        ((x / 10 + y + z * 10) / 5 > 116 and (x / 10 + y + z * 10) / 3 < 136)]:
        triggered.add(31)
    # 编号32 (原编号38)
    if [((x / 10 + y + z * 10) / 3 > 116 and (x / 10 + y + z * 10) / 3 < 136)] != [
        ((x / 10 + y + z * 10) / 3 > 116 and (x / 10 + y + z * 10) / 11 < 136)]:
        triggered.add(32)

    # -------------------------- 33-58：驱动系统分析 --------------------------
    # 编号33 (原编号41)
    if [(x / (y * 10) > 0.8 and x / (y * 10) < 1.0)] != [(x / (y * 10) > 8 and x / (y * 10) < 1.0)]:
        triggered.add(33)
    # 编号34 (原编号42)
    if [(x / (y * 10) > 0.8 and x / (y * 10) < 1.0)] != [(x / (y * 10) > 0.8 and x / (y * 10) < 10)]:
        triggered.add(34)
    # 编号35 (原编号43)
    if [(x / z > 110 and x / z < 140)] != [(x / z > 1220 and x / z < 140)]:
        triggered.add(35)
    # 编号36 (原编号44)
    if [(x / z > 110 and x / z < 140)] != [(x / z > 110 and x / z < 14)]:
        triggered.add(36)
    # 编号37 (原编号45)
    if [(y / z > 12 and y / z < 16)] != [(y / z > 112 and y / z < 16)]:
        triggered.add(37)
    # 编号38 (原编号46)
    if [(y / z > 12 and y / z < 16)] != [(y / z > 12 and y / z < 1226)]:
        triggered.add(38)
    # 编号39 (原编号47)
    if [((x - 1200) / 10 + (y - 150) > 35 and (x - 1200) / 10 + (y - 150) < 55)] != [
        ((x - 1200) / 10 + (y - 150) > 315 and (x - 1200) / 10 + (y - 150) < 55)]:
        triggered.add(39)
    # 编号40 (原编号48)
    if [((x - 1200) / 10 + (y - 150) > 35 and (x - 1200) / 10 + (y - 150) < 55)] != [
        ((x - 1200) / 10 + (y - 220) > 35 and (x - 1200) / 10 + (y - 150) < 55)]:
        triggered.add(40)
    # 编号41 (原编号49)
    if [((x - 1200) / 10 + (z - 10) * 5 > 40 and (x - 1200) / 10 + (z - 10) * 5 < 60)] != [
        ((x - 1200) / 10 + (z - 10) * 15 > 40 and (x - 1200) / 10 + (z - 10) * 5 < 60)]:
        triggered.add(41)
    # 编号42 (原编号50)
    if [((x - 1200) / 10 + (z - 10) * 5 > 40 and (x - 1200) / 10 + (z - 10) * 5 < 60)] != [
        ((x - 1200) / 10 + (z - 10) * 115 > 40 and (x - 1200) / 10 + (z - 10) * 5 < 60)]:
        triggered.add(42)
    # 编号43 (原编号51)
    if [((y - 150) + (z - 10) * 2 > 30 and (y - 150) + (z - 10) * 2 < 50)] != [
        ((y - 150) + (z - 10) * 2 > 302 and (y - 150) + (z - 10) * 2 < 50)]:
        triggered.add(43)
    # 编号44 (原编号52)
    if [((y - 150) + (z - 10) * 2 > 30 and (y - 150) + (z - 10) * 2 < 50)] != [
        ((y - 150) + (z - 10) * 211 > 30 and (y - 150) + (z - 10) * 2 < 50)]:
        triggered.add(44)
    # 编号45 (原编号53)
    if [(abs((x - 1500) / 10 - (y - 175)) < 12)] != [(abs((x - 1500) / 10 - (y - 175)) < 112)]:
        triggered.add(45)
    # 编号46 (原编号54)
    if [(abs((x - 1500) / 10 - (y - 175)) < 12)] != [(abs((x - 1500) / 101 - (y - 175)) < 12)]:
        triggered.add(46)
    # 编号47 (原编号55)
    if [(abs((x - 1500) / 100 - (z - 12.5)) < 3)] != [(abs((x - 1500) / 100 - (z - 12.5)) < 43)]:
        triggered.add(47)
    # 编号48 (原编号56)
    if [(abs((x - 1500) / 100 - (z - 12.5)) < 3)] != [(abs((x - 1500) / 100 - (z - 12.5)) < 13)]:
        triggered.add(48)
    # 编号49 (原编号57)
    if [(abs((y - 175) - (z - 12.5) * 12) < 8)] != [(abs((y - 175) - (z - 12.5) * 12) < 118)]:
        triggered.add(49)
    # 编号50 (原编号58)
    if [(abs((y - 175) - (z - 12.5) * 12) < 8)] != [(abs((y - 175) - (z - 12.5) * 232) < 8)]:
        triggered.add(50)
    # 编号51 (原编号60)
    if [(x / (y * 10 + 500) > 0.7 and x / (y * 10 + 500) < 0.9)] != [
        (x / (y * 10 + 500) > 337 and x / (y * 10 + 500) < 0.9)]:
        triggered.add(51)
    # 编号52 (原编号61)
    if [(y / (z + 5) > 9 and y / (z + 5) < 13)] != [(y / (z + 5) > 119 and y / (z + 5) < 13)]:
        triggered.add(52)
    # 编号53 (原编号62)
    if [(y / (z + 5) > 9 and y / (z + 5) < 13)] != [(y / (z + 5) > 9 and y / (z + 5) < 123)]:
        triggered.add(53)
    # 编号54 (原编号63)
    if [(z / (x / 150) > 1.1 and z / (x / 150) < 1.5)] != [(z / (x / 150) > 11 and z / (x / 150) < 1.5)]:
        triggered.add(54)
    # 编号55 (原编号64)
    if [(z / (x / 150) > 1.1 and z / (x / 150) < 1.5)] != [(z / (x / 150) > 1.1 and z / (x / 150) < 15)]:
        triggered.add(55)
    # 编号56 (原编号65)
    if [((x * y * z / 10000) > 28 and (x * y * z / 10000) < 38)] != [
        ((x * y * z / 10000) > 28 and (x * y * z / 10000) < 238)]:
        triggered.add(56)
    # 编号57 (原编号67)
    if [(x / 20 + y * 0.6 + z * 8 > 260 and x / 20 + y * 0.6 + z * 8 < 300)] != [
        (x / 20 + y * 0.6 + z * 10 > 260 and x / 20 + y * 0.6 + z * 8 < 300)]:
        triggered.add(57)
    # 编号58 (原编号68)
    if [(x / 20 + y * 0.6 + z * 8 > 260 and x / 20 + y * 0.6 + z * 8 < 300)] != [
        (x / 20 + y * 0.6 + z * 8 > 260 and x / 20 + y * 6 + z * 8 < 300)]:
        triggered.add(58)
    # 编号59 (原编号69)
    if [((x / 1500) ** 0.4 * (y / 175) ** 0.4 * (z / 12.5) ** 0.2 > 0.93)] != [
        ((x / 1500) ** 0.4 * (y / 175) ** 0.4 * (z / 12.5) ** 0.2 > 93)]:
        triggered.add(59)
    # 编号60 (原编号70)
    if [((x / 1500) ** 0.4 * (y / 175) ** 0.4 * (z / 12.5) ** 0.2 > 0.93)] != [
        ((x / 1500) ** 0.4 * (y / 175) ** 4 * (z / 12.5) ** 0.2 > 0.93)]:
        triggered.add(60)
    # 编号61 (原编号71)
    if [((x - 1500) / 10 * (y - 175) > -200 and (x - 1500) / 10 * (y - 175) < 200)] != [
        ((x - 1500) / 110 * (y - 175) > -200 and (x - 1500) / 10 * (y - 175) < 200)]:
        triggered.add(61)
    # 编号62 (原编号72)
    if [((x - 1500) / 10 * (y - 175) > -200 and (x - 1500) / 10 * (y - 175) < 200)] != [
        ((x - 1500) / 10 * (y - 175) > -200 and (x - 1500) / 110 * (y - 175) < 200)]:
        triggered.add(62)
    # 编号63 (原编号73)
    if [((x - 1500) / 100 * (z - 12.5) > -20 and (x - 1500) / 100 * (z - 12.5) < 20)] != [
        ((x - 1500) / 100 * (z - 12.5) > -20 and (x - 1500) / 10 * (z - 12.5) < 20)]:
        triggered.add(63)
    # 编号64 (原编号75)
    if [((y - 175) * (z - 12.5) > -20 and (y - 175) * (z - 12.5) < 20)] != [
        ((y - 175) * (z - 12.5) > -20 and (y - 175) * (z - 12.5) < 210)]:
        triggered.add(64)
    # 编号65 (原编号76)
    if [((y - 175) * (z - 12.5) > -20 and (y - 175) * (z - 12.5) < 20)] != [((y - 175) * (z - 125) < 20)]:
        triggered.add(65)
    # 编号66 (原编号77)
    if [(abs(x / 1500 + y / 175 + z / 12.5 - 3) < 0.2)] != [(abs(x / 1500 + y / 175 + z / 12.5 - 3) < 112)]:
        triggered.add(66)
    # 编号67 (原编号78)
    if [(abs(x / 1500 + y / 175 + z / 12.5 - 3) < 0.2)] != [(abs(x / 1500 + y / 175 + z / 12.5 - 3) < 2222)]:
        triggered.add(67)
    # 编号68 (原编号79)
    if [(abs((x / 1500) * (y / 175) * (z / 12.5) - 1) < 0.15)] != [
        (abs((x / 1500) * (y / 175) * (z / 12.5) - 1) < 115)]:
        triggered.add(68)
    # 编号69 (原编号80)
    if [(abs((x / 1500) * (y / 175) * (z / 12.5) - 1) < 0.15)] != [
        (abs((x / 1500) * (y / 175) * (z / 12.5) - 1) < 2225)]:
        triggered.add(69)

    # -------------------------- 70-107：动力传动协调 --------------------------
    # 编号70 (原编号81)
    if [(x > 1480 and x < 1520 and y > 172 and y < 178 and z > 12.2 and z < 12.8)] != [
        (x > 1480 and x < 1520 and y > 172 and y < 178 and z > 12.2 and 8 < 12.8)]:
        triggered.add(70)
    # 编号71 (原编号82)
    if [(x > 1480 and x < 1520 and y > 172 and y < 178 and z > 12.2 and z < 12.8)] != [
        (x > 1480 and x < 1520 and y > 172 and y < 178 and z > 12.2 and 10 < 12.8)]:
        triggered.add(71)
    # 编号72 (原编号83)
    if [(abs(x / (y * 10) - 0.86) < 0.06 and abs(z - 12.5) < 0.3)] != [
        (abs(x / (y * 10) - 0.86) < 0.06 and abs(z - 12.5) < 3)]:
        triggered.add(72)
    # 编号73 (原编号84)
    if [(abs(x / (y * 10) - 0.86) < 0.06 and abs(z - 12.5) < 0.3)] != [
        (abs(x / (y * 10) - 0.86) < 0.06 and abs(z - 12.5) < 13)]:
        triggered.add(73)
    # 编号74 (原编号85)
    if [(x * y / 1000 > 258 and x * y / 1000 < 268 and z > 12.3 and z < 12.7)] != [
        (x * y / 1000 > 258 and x * y / 1000 < 268 and 8 * z > 12.3 and z < 12.7)]:
        triggered.add(74)
    # 编号75 (原编号86)
    if [(x * y / 1000 > 258 and x * y / 1000 < 268 and z > 12.3 and z < 12.7)] != [
        (x * y / 1000 > 258 and x * y / 1000 < 268 and z * 99 > 12.3 and z < 12.7)]:
        triggered.add(75)
    # 编号76 (原编号87)
    if [((x / 10 + y) / 2 > 137 and (x / 10 + y) / 2 < 143 and z > 12.2 and z < 12.8)] != [
        ((x / 10 + y) / 2 > 137 and (x / 10 + y) / 2 < 143 and z > 12.2 and 8 < 12.8)]:
        triggered.add(76)
    # 编号77 (原编号88)
    if [((x / 10 + y) / 2 > 137 and (x / 10 + y) / 2 < 143 and z > 12.2 and z < 12.8)] != [
        ((x / 10 + y) / 2 > 137 and (x / 10 + y) / 2 < 143 and z > 12.2 and 10 < 12.8)]:
        triggered.add(77)
    # 编号78 (原编号89)
    if [(abs(x / 10 - y) < 25 and z > 12.2 and z < 12.8)] != [(abs(x / 10 - y) < 25 and z * 12 > 12.2 and z < 12.8)]:
        triggered.add(78)
    # 编号79 (原编号90)
    if [(abs(x / 10 - y) < 25 and z > 12.2 and z < 12.8)] != [(abs(x / 10 - y) < 25 and z * 89 > 12.2 and z < 12.8)]:
        triggered.add(79)
    # 编号80 (原编号91)
    if [(math.sqrt((x / 10 - 150) ** 2 + (y - 175) ** 2) < 6 and abs(z - 12.5) < 0.3)] != [
        (math.sqrt((x / 10 - 150) ** 2 + (y - 175) ** 2) < 6 and abs(z - 12.5) < 3)]:
        triggered.add(80)
    # 编号81 (原编号92)
    if [(math.sqrt((x / 10 - 150) ** 2 + (y - 175) ** 2) < 6 and abs(z - 12.5) < 0.3)] != [
        (math.sqrt((x / 10 - 150) ** 2 + (y - 175) ** 2) < 6 and abs(z - 12.5) < 13)]:
        triggered.add(81)
    # 编号82 (原编号93)
    if [(x / 1500 > 0.98 and x / 1500 < 1.02 and y / 175 > 0.97 and y / 175 < 1.03)] != [
        (x / 1500 > 0.98 and x / 1500 < 1.02 and y / 175 > 97 and y / 175 < 1.03)]:
        triggered.add(82)
    # 编号83 (原编号94)
    if [(x / 1500 > 0.98 and x / 1500 < 1.02 and y / 175 > 0.97 and y / 175 < 1.03)] != [
        (x / 1500 > 0.98 and x / 1500 < 1.02 and y * 8 / 175 > 0.97 and y / 175 < 1.03)]:
        triggered.add(83)
    # 编号84 (原编号95)
    if [(z / 12.5 > 0.976 and z / 12.5 < 1.024)] != [(z / 12.5 > 0.976 and z / 125 < 1.024)]:
        triggered.add(84)
    # 编号85 (原编号96)
    if [(z / 12.5 > 0.976 and z / 12.5 < 1.024)] != [(z / 12.5 > 0.976 and z / 200 < 1.024)]:
        triggered.add(85)
    # 编号86 (原编号97)
    if [((x / 1500 + y / 175 + z / 12.5) / 3 > 0.98 and (x / 1500 + y / 175 + z / 12.5) / 3 < 1.02)] != [
        ((x / 1500 + y / 175 + z / 12.5) / 3 > 98 and (x / 1500 + y / 175 + z / 12.5) / 3 < 1.02)]:
        triggered.add(86)
    # 编号87 (原编号98)
    if [((x / 1500 + y / 175 + z / 12.5) / 3 > 0.98 and (x / 1500 + y / 175 + z / 12.5) / 3 < 1.02)] != [
        ((x / 1500 + y / 175 + z / 12.5) / 23 > 0.98 and (x / 1500 + y / 175 + z / 12.5) / 3 < 1.02)]:
        triggered.add(87)
    # 编号88 (原编号99)
    if [(max(abs(x / 1500 - 1), abs(y / 175 - 1), abs(z / 12.5 - 1)) < 0.025)] != [
        (max(abs(x / 1500 - 1), abs(y / 175 - 1), abs(z / 12.5 - 1)) < 0.55)]:
        triggered.add(88)
    # 编号89 (原编号100)
    if [(max(abs(x / 1500 - 1), abs(y / 175 - 1), abs(z / 12.5 - 1)) < 0.025)] != [
        (max(abs(x / 1500 - 1), abs(y / 175 - 1), abs(z / 12.5 - 1)) < 0.25)]:
        triggered.add(89)
    # 编号90 (原编号101)
    if [(min(x / 1500, y / 175, z / 12.5) > 0.98)] != [(min(x / 1500, y / 175, z / 12.5) > 498)]:
        triggered.add(90)
    # 编号91 (原编号102)
    if [(min(x / 1500, y / 175, z / 12.5) > 0.98)] != [(min(x / 1500, y / 1375, z / 12.5) > 0.98)]:
        triggered.add(91)
    # 编号92 (原编号103)
    if [(max(x / 1500, y / 175, z / 12.5) < 1.02)] != [(max(x / 1500, y / 175, z / 12.5) < 302)]:
        triggered.add(92)
    # 编号93 (原编号104)
    if [(max(x / 1500, y / 175, z / 12.5) < 1.02)] != [(max(x / 1500, y / 175, z / 12.5) < 102)]:
        triggered.add(93)
    # 编号94 (原编号105)
    if [(abs(max(x / 10, y, z * 10) - min(x / 10, y, z * 10)) < 30)] != [
        (abs(max(x / 10, y, z * 120) - min(x / 10, y, z * 10)) < 30)]:
        triggered.add(94)
    # 编号95 (原编号106)
    if [(abs(max(x / 10, y, z * 10) - min(x / 10, y, z * 10)) < 30)] != [
        (abs(max(x / 10, y, z * 10) - min(x / 10, y, z * 10)) < 230)]:
        triggered.add(95)
    # 编号96 (原编号107)
    if [((x / 10 + y + z * 10) > 378 and (x / 10 + y + z * 10) < 382)] != [
        ((x / 10 + y + z * 10) > 378 and (x / 10 + y + z * 10) < 3112)]:
        triggered.add(96)
    # 编号97 (原编号108)
    if [((x / 10 + y + z * 10) > 378 and (x / 10 + y + z * 10) < 382)] != [
        ((x / 10 + y + z * 10) > 3781 and (x / 10 + y + z * 10) < 382)]:
        triggered.add(97)
    # 编号98 (原编号109)
    if [((x * y * z / 10000) > 32.5 and (x * y * z / 10000) < 33.5)] != [
        ((x * y * z / 10000) > 32.5 and (x * y * 2 / 10000) < 33.5)]:
        triggered.add(98)
    # 编号99 (原编号110)
    if [((x * y * z / 10000) > 32.5 and (x * y * z / 10000) < 33.5)] != [
        ((x * y * z / 10000) > 32.5 and (x * y * 2.4 / 10000) < 33.5)]:
        triggered.add(99)
    # 编号100 (原编号111)
    if [(abs((x / 10 + y + z * 10) / 3 - 126.7) < 1.5)] != [(abs((x / 10 + y + z * 10) / 3 - 126.7) < 15)]:
        triggered.add(100)
    # 编号101 (原编号112)
    if [(abs((x / 10 + y + z * 10) / 3 - 126.7) < 1.5)] != [(abs((x / 10 + y + z * 10) / 3 - 126.7) < 115)]:
        triggered.add(101)
    # 编号102 (原编号113)
    if [(x > 1495 and x < 1505 and y > 174 and y < 176)] != [(x > 1495 and x < 1505 and y * 10 > 174 and y < 176)]:
        triggered.add(102)
    # 编号103 (原编号114)
    if [(x > 1495 and x < 1505 and y > 174 and y < 176)] != [(x > 1495 and x < 1505 and y * 88 > 174 and y < 176)]:
        triggered.add(103)
    # 编号104 (原编号115)
    if [(z > 12.45 and z < 12.55)] != [(z * 10 > 12.45 and z < 12.55)]:
        triggered.add(104)
    # 编号105 (原编号116)
    if [(z > 12.45 and z < 12.55)] != [(z * 8 > 12.45 and z < 12.55)]:
        triggered.add(105)

    # -------------------------- 106-144：动力控制优化 --------------------------
    # 编号106 (原编号117)
    if [(x < 1350 or y < 160 or z < 11.5)] != [(x < 1350 or y * 8 < 160 or z < 11.5)]:
        triggered.add(106)
    # 编号107 (原编号118)
    if [(x < 1350 or y < 160 or z < 11.5)] != [(x < 1350 or y < 160 or z * 89 < 11.5)]:
        triggered.add(107)
    # 编号108 (原编号119)
    if [(x > 1650 or y > 190 or z > 13.5)] != [(x > 1650 or y > 190 or z > 135)]:
        triggered.add(108)
    # 编号109 (原编号120)
    if [(x > 1650 or y > 190 or z > 13.5)] != [(x > 1650 or y * 87 > 190 or z > 13.5)]:
        triggered.add(109)
    # 编号110 (原编号121)
    if [(abs(x / (y * 10) - 0.86) > 0.1)] != [(abs(x / (y * 10) - 0.86) > 111)]:
        triggered.add(110)
    # 编号111 (原编号122)
    if [(abs(x / (y * 10) - 0.86) > 0.1)] != [(abs(x / (y * 10) - 0.86) > 1671)]:
        triggered.add(111)
    # 编号112 (原编号123)
    if [(abs(x / z - 120) > 15)] != [(abs(x / z - 120) > 1522)]:
        triggered.add(112)
    # 编号113 (原编号124)
    if [(abs(x / z - 120) > 15)] != [(abs(x / z - 120) > 115)]:
        triggered.add(113)
    # 编号114 (原编号125)
    if [(abs(y / z - 14) > 2)] != [(abs(y / z - 14) > 112)]:
        triggered.add(114)
    # 编号115 (原编号126)
    if [(abs(y / z - 14) > 2)] != [(abs(y / z - 14) > 211)]:
        triggered.add(115)
    # 编号116 (原编号127)
    if [((x / 10 + y + z * 10) < 360 or (x / 10 + y + z * 10) > 400)] != [
        ((x / 10 + y + z * 10) < 36 or (x / 10 + y + z * 10) > 400)]:
        triggered.add(116)
    # 编号117 (原编号128)
    if [((x / 10 + y + z * 10) < 360 or (x / 10 + y + z * 10) > 400)] != [
        ((x / 10 + y + z * 10) < 989 or (x / 10 + y + z * 10) > 400)]:
        triggered.add(117)
    # 编号118 (原编号129)
    if [((x * y * z / 10000) < 30 or (x * y * z / 10000) > 36)] != [
        ((x * y * z / 10000) < 30 or (x * y * z / 10000) > 136)]:
        triggered.add(118)
    # 编号119 (原编号130)
    if [((x * y * z / 10000) < 30 or (x * y * z / 10000) > 36)] != [
        ((x * y * z / 10000) < 30 or (x * y * z / 10000) > 316)]:
        triggered.add(119)
    # 编号120 (原编号131)
    if [(x < 1300 and y < 160)] != [(x < 1300 and y * 878 < 160)]:
        triggered.add(120)
    # 编号121 (原编号132)
    if [(x < 1300 and y < 160)] != [(x * 71 < 1300 and y < 160)]:
        triggered.add(121)
    # 编号122 (原编号133)
    if [(x > 1700 and y > 190)] != [(x > 1700 and y * 78 > 190)]:
        triggered.add(122)
    # 编号123 (原编号134)
    if [(x > 1700 and y > 190)] != [(x * 78 > 1700 and y > 190)]:
        triggered.add(123)
    # 编号124 (原编号135)
    if [(x < 1300 and z < 11)] != [(x < 1300 and z * 91 < 11)]:
        triggered.add(124)
    # 编号125 (原编号136)
    if [(x < 1300 and z < 11)] != [(x * 12 < 1300 and z < 11)]:
        triggered.add(125)
    # 编号126 (原编号137)
    if [(x > 1700 and z > 14)] != [(x > 1700 and z * 21 > 14)]:
        triggered.add(126)
    # 编号127 (原编号138)
    if [(x > 1700 and z > 14)] != [(x * 123 > 1700 and z > 14)]:
        triggered.add(127)
    # 编号128 (原编号139)
    if [(y < 160 and z < 11)] != [(y < 160 and z * 78 < 11)]:
        triggered.add(128)
    # 编号129 (原编号140)
    if [(y < 160 and z < 11)] != [(y * 8 < 160 and z < 11)]:
        triggered.add(129)
    # 编号130 (原编号141)
    if [(y > 190 and z > 14)] != [(y * 8 > 190 and z > 14)]:
        triggered.add(130)
    # 编号131 (原编号142)
    if [(y > 190 and z > 14)] != [(y > 190 and z * 9 > 14)]:
        triggered.add(131)
    # 编号132 (原编号143)
    if [(x < 1250 or x > 1750)] != [(x * 67 < 1250 or x > 1750)]:
        triggered.add(132)
    # 编号133 (原编号144)
    if [(x < 1250 or x > 1750)] != [(x < 1250 or x * 53 > 1750)]:
        triggered.add(133)
    # 编号134 (原编号145)
    if [(y < 145 or y > 205)] != [(y * 67 < 145 or y > 205)]:
        triggered.add(134)
    # 编号135 (原编号146)
    if [(y < 145 or y > 205)] != [(y < 145 or y * 67 > 205)]:
        triggered.add(135)
    # 编号136 (原编号147)
    if [(z < 9.5 or z > 15.5)] != [(z < 9.5 or z * 67 > 15.5)]:
        triggered.add(136)
    # 编号137 (原编号149)
    if [(x < 1200 and y < 155 and z < 10.5)] != [(x * 67 < 1200 and y < 155 and z < 10.5)]:
        triggered.add(137)
    # 编号138 (原编号150)
    if [(x < 1200 and y < 155 and z < 10.5)] != [(x < 1200 and y < 55 and z < 10.5)]:
        triggered.add(138)
    # 编号139 (原编号151)
    if [(x > 1800 and y > 195 and z > 14.5)] != [(x > 1800 and y * 78 > 195 and z > 14.5)]:
        triggered.add(139)
    # 编号140 (原编号152)
    if [(x > 1800 and y > 195 and z > 14.5)] != [(x > 1800 and y > 966 and z > 14.5)]:
        triggered.add(140)
    # 编号141 (原编号153)
    if [((x / 10 - 150) ** 2 + (y - 175) ** 2 + (z - 12.5) ** 2 * 100) > 150] != [
        ((x / 10 - 150) ** 6 + (y - 175) ** 2 + (z - 12.5) ** 2 * 100) > 150]:
        triggered.add(141)
    # 编号142 (原编号154)
    if [((x / 10 - 150) ** 2 + (y - 175) ** 2 + (z - 12.5) ** 2 * 100) > 150] != [
        ((x / 10 - 150) ** 2 + (y - 175) ** 8 + (z - 12.5) ** 2 * 100) > 150]:
        triggered.add(142)
    # 编号143 (原编号155)
    if [(abs((x * y * z / 10000) / 33 - 1) > 0.18)] != [(abs((x * y * z / 10000) / 33 - 1) > 18)]:
        triggered.add(143)
    # 编号144 (原编号156)
    if [(abs((x * y * z / 10000) / 33 - 1) > 0.18)] != [(abs((x * y * z / 10000) / 33 - 1) > 10)]:
        triggered.add(144)

    return triggered

# ================================================================
# 3. 别名指向 section8
# ================================================================
execute_Tr = section8_hybrid_speed_torque_current

# ================================================================
# 4. Jaccard 相似度
# ================================================================
def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0

# ================================================================
# 5. 目标路径组（重要！请根据实际测试需求替换）
#    这里仅给出占位示例，包含部分规则编号。
#    您必须将下面的集合替换为真正关心的规则组合。
# ================================================================

# === 目标路径组 ===
targetPaths = [
    {1, 4, 6, 9, 11, 12, 13, 14, 15, 16, 17, 21, 22, 24, 25, 27, 29, 32, 33, 38, 44, 45, 46, 47, 48, 49, 51, 52, 58, 59,
     60, 64, 68, 69, 74, 75, 78, 79, 88, 89, 92, 93, 95, 96, 101, 104, 105, 107, 109, 110, 111, 112, 113, 114, 115, 119,
     133, 135, 136, 143, 144},

    {4, 6, 9, 11, 12, 13, 14, 15, 16, 17, 21, 22, 24, 25, 27, 29, 32, 34, 37, 39, 40, 44, 45, 47, 48, 50, 51, 52, 58,
     59, 60, 61, 64, 68, 69, 74, 75, 78, 79, 88, 89, 92, 93, 95, 96, 101, 104, 105, 109, 110, 111, 112, 113, 119, 128,
     129, 133, 135, 136, 143, 144},

    {3, 5, 9, 11, 12, 13, 14, 16, 17, 18, 21, 22, 24, 25, 27, 29, 32, 34, 35, 36, 37, 39, 40, 44, 45, 49, 51, 52, 54,
     58, 59, 60, 61, 65, 74, 75, 78, 79, 86, 87, 88, 89, 92, 93, 95, 96, 101, 104, 105, 106, 109, 110, 111, 112, 113,
     119, 133, 135, 136, 143, 144},

    {2, 3, 8, 9, 11, 12, 13, 14, 16, 17, 18, 19, 21, 22, 23, 27, 29, 31, 34, 35, 36, 41, 42, 44, 45, 49, 51, 54, 56, 57,
     62, 64, 66, 67, 68, 69, 78, 79, 88, 89, 94, 96, 100, 101, 104, 105, 106, 109, 110, 111, 114, 115, 117, 119, 133,
     134, 136, 143, 144},

    {2, 4, 7, 9, 11, 12, 16, 17, 21, 22, 24, 27, 29, 32, 33, 35, 36, 37, 41, 42, 44, 45, 50, 51, 52, 57, 64, 66, 67, 68,
     69, 78, 79, 83, 88, 89, 95, 96, 100, 101, 102, 103, 104, 105, 109, 110, 111, 112, 113, 119, 128, 129, 133, 135,
     136, 143, 144},

    {6, 8, 11, 12, 14, 15, 16, 17, 21, 22, 24, 29, 32, 35, 36, 38, 39, 40, 42, 43, 45, 49, 52, 54, 56, 57, 60, 61, 63,
     65, 66, 67, 68, 69, 88, 89, 92, 93, 95, 96, 100, 101, 104, 105, 109, 110, 111, 114, 115, 119, 124, 125, 132, 135,
     136, 143, 144},

    {1, 4, 6, 8, 9, 11, 12, 13, 14, 15, 16, 17, 21, 22, 24, 25, 29, 32, 33, 35, 36, 37, 39, 40, 42, 44, 49, 52, 54, 57,
     59, 60, 65, 66, 67, 68, 69, 72, 73, 78, 79, 88, 89, 95, 96, 100, 101, 104, 105, 107, 109, 119, 133, 135, 136, 143,
     144},

    {1, 3, 5, 7, 9, 11, 12, 17, 18, 21, 22, 24, 25, 27, 30, 32, 33, 35, 36, 37, 39, 40, 41, 42, 44, 50, 52, 54, 58, 59,
     65, 72, 73, 74, 75, 78, 79, 80, 81, 82, 88, 89, 95, 96, 101, 102, 103, 104, 105, 109, 119, 133, 135, 136, 142, 143,
     144},

    {8, 9, 11, 12, 14, 16, 17, 21, 22, 23, 29, 31, 33, 35, 36, 37, 44, 46, 49, 54, 56, 62, 63, 64, 66, 67, 68, 69, 72,
     73, 78, 79, 88, 89, 95, 98, 100, 101, 104, 105, 109, 117, 119, 120, 121, 124, 125, 128, 129, 132, 134, 136, 137,
     138, 143, 144},

    {9, 11, 12, 14, 16, 17, 24, 25, 27, 29, 32, 34, 37, 39, 40, 45, 47, 48, 50, 51, 52, 57, 59, 60, 61, 64, 68, 69, 74,
     75, 78, 79, 88, 89, 92, 93, 95, 96, 100, 101, 104, 105, 110, 111, 112, 113, 119, 122, 126, 128, 129, 133, 135, 136,
     143, 144},

    {6, 8, 11, 12, 14, 16, 17, 21, 22, 23, 29, 31, 35, 36, 38, 39, 40, 43, 45, 49, 52, 54, 56, 57, 60, 61, 63, 65, 66,
     67, 68, 69, 88, 89, 92, 93, 95, 96, 100, 101, 104, 105, 110, 111, 114, 115, 119, 123, 124, 125, 131, 132, 135, 136,
     143, 144},

    {8, 9, 11, 12, 14, 16, 17, 21, 22, 29, 34, 35, 36, 42, 44, 45, 49, 51, 54, 56, 62, 63, 64, 66, 67, 68, 69, 78, 79,
     88, 94, 98, 99, 100, 101, 104, 105, 109, 110, 111, 114, 115, 116, 119, 120, 121, 124, 125, 128, 129, 133, 134, 136,
     143, 144},

    {5, 8, 11, 12, 14, 16, 17, 20, 21, 22, 24, 30, 32, 38, 39, 40, 42, 43, 45, 49, 52, 54, 58, 60, 61, 63, 65, 66, 67,
     68, 69, 88, 89, 92, 93, 95, 96, 100, 101, 104, 105, 110, 111, 112, 113, 114, 115, 119, 123, 131, 132, 135, 136,
     143, 144},

    {8, 9, 14, 16, 17, 23, 27, 29, 31, 34, 41, 42, 44, 45, 47, 48, 49, 51, 55, 56, 57, 62, 66, 67, 68, 69, 84, 85, 88,
     92, 93, 95, 96, 98, 99, 100, 101, 108, 110, 111, 112, 113, 114, 115, 117, 119, 120, 121, 127, 130, 132, 134, 136,
     143, 144},

    {5, 8, 9, 14, 15, 16, 17, 18, 20, 21, 22, 23, 30, 32, 37, 41, 42, 44, 45, 46, 47, 48, 49, 52, 55, 58, 62, 65, 66,
     67, 68, 69, 76, 77, 84, 85, 88, 89, 92, 93, 95, 96, 100, 101, 109, 110, 111, 112, 113, 119, 132, 135, 136, 143,
     144},

    {1, 3, 5, 9, 11, 12, 17, 18, 21, 22, 24, 25, 27, 30, 32, 33, 35, 36, 37, 39, 40, 44, 50, 52, 54, 58, 59, 60, 65, 72,
     73, 74, 75, 78, 79, 80, 81, 82, 88, 89, 95, 96, 101, 104, 105, 109, 119, 133, 135, 136, 141, 142, 143, 144},

    {1, 4, 6, 8, 10, 11, 12, 17, 21, 22, 24, 25, 29, 32, 33, 35, 36, 38, 39, 40, 42, 44, 49, 52, 58, 59, 68, 69, 72, 73,
     74, 75, 80, 81, 82, 88, 89, 95, 96, 101, 104, 105, 107, 109, 112, 113, 114, 115, 119, 133, 135, 136, 143, 144},

    {8, 9, 11, 12, 14, 16, 17, 21, 22, 23, 29, 31, 33, 35, 36, 37, 42, 44, 46, 50, 52, 54, 56, 62, 63, 64, 66, 67, 68,
     69, 72, 73, 78, 79, 88, 89, 95, 97, 104, 105, 109, 117, 119, 120, 121, 124, 125, 128, 129, 133, 135, 136, 143,
     144},

    {8, 9, 11, 12, 14, 16, 17, 21, 22, 29, 34, 42, 44, 45, 49, 51, 54, 56, 63, 64, 66, 67, 68, 69, 78, 79, 88, 94, 98,
     99, 101, 104, 105, 109, 110, 111, 112, 113, 114, 115, 116, 118, 119, 120, 121, 124, 125, 128, 129, 132, 134, 136,
     143, 144},

    {5, 11, 12, 14, 15, 16, 17, 24, 26, 27, 30, 32, 33, 38, 43, 45, 46, 47, 48, 49, 51, 52, 58, 59, 62, 66, 67, 68, 69,
     78, 79, 88, 89, 92, 93, 95, 96, 101, 104, 105, 107, 110, 111, 112, 113, 114, 115, 122, 126, 135, 136, 143},

    {1, 3, 5, 7, 9, 21, 22, 24, 25, 27, 30, 32, 33, 35, 36, 37, 39, 40, 44, 50, 52, 54, 58, 59, 65, 70, 71, 72, 73, 80,
     81, 82, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 95, 96, 101, 109, 133, 135, 136, 142, 143, 144},

    {2, 8, 9, 13, 14, 16, 17, 19, 24, 28, 30, 32, 34, 44, 45, 47, 48, 49, 51, 55, 58, 59, 60, 62, 76, 77, 84, 85, 86,
     87, 88, 92, 93, 94, 96, 101, 106, 108, 110, 111, 112, 113, 114, 115, 119, 127, 130, 133, 134, 143, 144},

    {14, 15, 16, 17, 24, 26, 28, 32, 33, 35, 36, 37, 43, 45, 46, 49, 51, 52, 54, 59, 62, 63, 64, 65, 66, 67, 68, 69, 84,
     85, 88, 89, 90, 91, 92, 93, 95, 96, 101, 110, 111, 122, 130, 135, 136, 139, 143},

    {11, 12, 14, 16, 17, 24, 26, 27, 29, 32, 33, 38, 43, 46, 47, 48, 49, 51, 53, 58, 59, 62, 65, 72, 73, 78, 79, 88, 89,
     92, 93, 95, 96, 101, 104, 105, 107, 112, 113, 114, 115, 126, 131, 135, 136, 143, 144},

    {14, 16, 17, 24, 26, 28, 32, 33, 35, 36, 37, 45, 49, 51, 52, 54, 59, 62, 63, 64, 65, 66, 67, 68, 69, 84, 85, 88, 90,
     91, 92, 93, 95, 96, 101, 110, 111, 135, 136, 140, 143},

]

# ================================================================
# 6. 实验配置
# ================================================================
class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 2000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(targetPaths)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }

# ================================================================
# 7. 鲁棒性计算（注意邻域步长，对于连续值可调整）
# ================================================================
def compute_robustness(state, path):
    base = execute_Tr(*state)
    if not base:
        return 0.0
    rob, neighbors = 0.0, 0
    # 邻域步长：对于速度、扭矩、电流，步长 1 可能导致变化过小，可适当放大
    # 这里保留原步长 1，您可根据需要修改为 step = [10, 5, 0.5] 等
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                    np.clip(state[0] + dx, STATE_MIN_X, STATE_MAX_X),
                    np.clip(state[1] + dy, STATE_MIN_Y, STATE_MAX_Y),
                    np.clip(state[2] + dz, STATE_MIN_Z, STATE_MAX_Z)
                ])
                n_trig = execute_Tr(*neighbor)
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0

# ================================================================
# 8. 候选样本生成（使用 uniform 采样）
# ================================================================
def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = targetPaths[target_path_idx]
    samples = []
    attempts = 0
    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.uniform(STATE_MIN_X, STATE_MAX_X),
            random.uniform(STATE_MIN_Y, STATE_MAX_Y),
            random.uniform(STATE_MIN_Z, STATE_MAX_Z)
        ])
        triggered = execute_Tr(*state)
        if not triggered:
            continue
        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)
        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }
        samples.append(sample_data)
    return samples

def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample['robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)
        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]
    return selected_samples

def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = targetPaths[target_path_idx]
        samples = []
        attempts = 0
        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.uniform(STATE_MIN_X, STATE_MAX_X),
                random.uniform(STATE_MIN_Y, STATE_MAX_Y),
                random.uniform(STATE_MIN_Z, STATE_MAX_Z)
            ])
            triggered = execute_Tr(*state)
            if not triggered:
                continue
            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)
            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)
        random.shuffle(samples)
        return samples[:config.top_k_samples]
    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")
        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)

# ================================================================
# 9. 单次实验
# ================================================================
def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}
    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates
    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}
        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config, shared_candidates[path_idx])
            strategy_results[path_idx] = samples
        results[strategy_name] = strategy_results
    return results

# ================================================================
# 10. 结果分析
# ================================================================
def analyze_fitness_values(results, config):
    analysis_results = {}
    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []
        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])
        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }
        all_scores = []
        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])
        scores_array = np.array(all_scores)
        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)
        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })
        analysis_results[strategy_name] = analysis
    return analysis_results

def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []
    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)
    df = pd.DataFrame(df_data)
    return df, analysis_results

# ================================================================
# 11. 多轮实验
# ================================================================
def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []
    print(f"Starting {num_runs} experiments...")
    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")
        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)
        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)
        print(f"Experiment {run_idx + 1} completed.")
    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df

# ================================================================
# 12. 保存结果到 Excel
# ================================================================
def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "Raw Data"
    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')
    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width
    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []
    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]
        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)
    stats_df = pd.DataFrame(stats_data)
    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)
    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width
    wb.save(output_path)
    print(f"Results saved to: {output_path}")

# ================================================================
# 13. 主函数
# ================================================================
def main():
    results_df = run_multiple_experiments(num_runs=20)
    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")
    save_results_to_excel(results_df, output_path)
    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)
    return results_df, output_path

if __name__ == "__main__":
    results, output_file = main()