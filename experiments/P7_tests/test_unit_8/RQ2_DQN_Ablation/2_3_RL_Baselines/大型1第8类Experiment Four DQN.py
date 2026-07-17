import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
import numpy as np
import random
import time
import math
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# === 配置（新范围，匹配 section8） ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'MIN_X': 1000, 'MAX_X': 2000,    # 速度范围 1000~2000
    'MIN_Y': 100, 'MAX_Y': 200,      # 扭矩范围 100~200
    'MIN_Z': 10, 'MAX_Z': 30,        # 电流范围 10~30
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 5,
    'NUM_ROUNDS': 5,
    'REPLAY_BATCH_SIZE': 64,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'LEARNING_RATE': 3e-4,
    'NUM_RUNS': 20,
    'TOP_K_SAMPLES': 20,
    'REPLAY_BUFFER_CAPACITY': 20000,
}

# === 目标路径组（与 section8 规则编号 1~144 匹配） ===
targetPaths = [
    {1, 4, 6, 9, 11, 12, 13, 14, 15, 16, 17, 21, 22, 24, 25, 27, 29, 32, 33, 38, 44, 45, 46, 47, 48, 49, 51, 52, 58, 59,
     60, 64, 68, 69, 74, 75, 78, 79, 88, 89, 92, 93, 95, 96, 101, 104, 105, 107, 109, 110, 111, 112, 113, 114, 115, 119,
     133, 135, 136, 143, 144},

    {4, 6, 9, 11, 12, 13, 14, 15, 16, 17, 21, 22, 24, 25, 27, 29, 32, 34, 37, 39, 40, 44, 45, 47, 48, 50, 51, 52, 58,
     59, 60, 61, 64, 68, 69, 74, 75, 78, 79, 88, 89, 92, 93, 95, 96, 101, 104, 105, 109, 110, 111, 112, 113, 119, 128,
     129, 133, 135, 136, 143, 144},

    {3, 5, 9, 11, 12, 13, 14, 16, 17, 18, 21, 22, 24, 25, 27, 29, 32, 34, 35, 36, 37, 39, 40, 44, 45, 49, 51, 52, 54,
     58, 59, 60, 61, 65, 74, 75, 78, 79, 86, 87, 88, 89, 92, 93, 95, 96, 101, 104, 105, 106, 109, 110, 111, 112, 113,
     119, 133, 135, 136, 143, 144},

    {2, 3, 8, 9, 11, 12, 13, 14, 16, 17, 18, 19, 21, 22, 23, 27, 29, 31, 34, 35, 36, 41, 42, 44, 45, 49, 51, 54, 56, 57,
     62, 64, 66, 67, 68, 69, 78, 79, 88, 89, 94, 96, 100, 101, 104, 105, 106, 109, 110, 111, 114, 115, 117, 119, 133,
     134, 136, 143, 144},

    {2, 4, 7, 9, 11, 12, 16, 17, 21, 22, 24, 27, 29, 32, 33, 35, 36, 37, 41, 42, 44, 45, 50, 51, 52, 57, 64, 66, 67, 68,
     69, 78, 79, 83, 88, 89, 95, 96, 100, 101, 102, 103, 104, 105, 109, 110, 111, 112, 113, 119, 128, 129, 133, 135,
     136, 143, 144},

    {6, 8, 11, 12, 14, 15, 16, 17, 21, 22, 24, 29, 32, 35, 36, 38, 39, 40, 42, 43, 45, 49, 52, 54, 56, 57, 60, 61, 63,
     65, 66, 67, 68, 69, 88, 89, 92, 93, 95, 96, 100, 101, 104, 105, 109, 110, 111, 114, 115, 119, 124, 125, 132, 135,
     136, 143, 144},

    {1, 4, 6, 8, 9, 11, 12, 13, 14, 15, 16, 17, 21, 22, 24, 25, 29, 32, 33, 35, 36, 37, 39, 40, 42, 44, 49, 52, 54, 57,
     59, 60, 65, 66, 67, 68, 69, 72, 73, 78, 79, 88, 89, 95, 96, 100, 101, 104, 105, 107, 109, 119, 133, 135, 136, 143,
     144},

    {1, 3, 5, 7, 9, 11, 12, 17, 18, 21, 22, 24, 25, 27, 30, 32, 33, 35, 36, 37, 39, 40, 41, 42, 44, 50, 52, 54, 58, 59,
     65, 72, 73, 74, 75, 78, 79, 80, 81, 82, 88, 89, 95, 96, 101, 102, 103, 104, 105, 109, 119, 133, 135, 136, 142, 143,
     144},

    {8, 9, 11, 12, 14, 16, 17, 21, 22, 23, 29, 31, 33, 35, 36, 37, 44, 46, 49, 54, 56, 62, 63, 64, 66, 67, 68, 69, 72,
     73, 78, 79, 88, 89, 95, 98, 100, 101, 104, 105, 109, 117, 119, 120, 121, 124, 125, 128, 129, 132, 134, 136, 137,
     138, 143, 144},

    {9, 11, 12, 14, 16, 17, 24, 25, 27, 29, 32, 34, 37, 39, 40, 45, 47, 48, 50, 51, 52, 57, 59, 60, 61, 64, 68, 69, 74,
     75, 78, 79, 88, 89, 92, 93, 95, 96, 100, 101, 104, 105, 110, 111, 112, 113, 119, 122, 126, 128, 129, 133, 135, 136,
     143, 144},

    {6, 8, 11, 12, 14, 16, 17, 21, 22, 23, 29, 31, 35, 36, 38, 39, 40, 43, 45, 49, 52, 54, 56, 57, 60, 61, 63, 65, 66,
     67, 68, 69, 88, 89, 92, 93, 95, 96, 100, 101, 104, 105, 110, 111, 114, 115, 119, 123, 124, 125, 131, 132, 135, 136,
     143, 144},

    {8, 9, 11, 12, 14, 16, 17, 21, 22, 29, 34, 35, 36, 42, 44, 45, 49, 51, 54, 56, 62, 63, 64, 66, 67, 68, 69, 78, 79,
     88, 94, 98, 99, 100, 101, 104, 105, 109, 110, 111, 114, 115, 116, 119, 120, 121, 124, 125, 128, 129, 133, 134, 136,
     143, 144},

    {5, 8, 11, 12, 14, 16, 17, 20, 21, 22, 24, 30, 32, 38, 39, 40, 42, 43, 45, 49, 52, 54, 58, 60, 61, 63, 65, 66, 67,
     68, 69, 88, 89, 92, 93, 95, 96, 100, 101, 104, 105, 110, 111, 112, 113, 114, 115, 119, 123, 131, 132, 135, 136,
     143, 144},

    {8, 9, 14, 16, 17, 23, 27, 29, 31, 34, 41, 42, 44, 45, 47, 48, 49, 51, 55, 56, 57, 62, 66, 67, 68, 69, 84, 85, 88,
     92, 93, 95, 96, 98, 99, 100, 101, 108, 110, 111, 112, 113, 114, 115, 117, 119, 120, 121, 127, 130, 132, 134, 136,
     143, 144},

    {5, 8, 9, 14, 15, 16, 17, 18, 20, 21, 22, 23, 30, 32, 37, 41, 42, 44, 45, 46, 47, 48, 49, 52, 55, 58, 62, 65, 66,
     67, 68, 69, 76, 77, 84, 85, 88, 89, 92, 93, 95, 96, 100, 101, 109, 110, 111, 112, 113, 119, 132, 135, 136, 143,
     144},

    {1, 3, 5, 9, 11, 12, 17, 18, 21, 22, 24, 25, 27, 30, 32, 33, 35, 36, 37, 39, 40, 44, 50, 52, 54, 58, 59, 60, 65, 72,
     73, 74, 75, 78, 79, 80, 81, 82, 88, 89, 95, 96, 101, 104, 105, 109, 119, 133, 135, 136, 141, 142, 143, 144},

    {1, 4, 6, 8, 10, 11, 12, 17, 21, 22, 24, 25, 29, 32, 33, 35, 36, 38, 39, 40, 42, 44, 49, 52, 58, 59, 68, 69, 72, 73,
     74, 75, 80, 81, 82, 88, 89, 95, 96, 101, 104, 105, 107, 109, 112, 113, 114, 115, 119, 133, 135, 136, 143, 144},

    {8, 9, 11, 12, 14, 16, 17, 21, 22, 23, 29, 31, 33, 35, 36, 37, 42, 44, 46, 50, 52, 54, 56, 62, 63, 64, 66, 67, 68,
     69, 72, 73, 78, 79, 88, 89, 95, 97, 104, 105, 109, 117, 119, 120, 121, 124, 125, 128, 129, 133, 135, 136, 143,
     144},

    {8, 9, 11, 12, 14, 16, 17, 21, 22, 29, 34, 42, 44, 45, 49, 51, 54, 56, 63, 64, 66, 67, 68, 69, 78, 79, 88, 94, 98,
     99, 101, 104, 105, 109, 110, 111, 112, 113, 114, 115, 116, 118, 119, 120, 121, 124, 125, 128, 129, 132, 134, 136,
     143, 144},

    {5, 11, 12, 14, 15, 16, 17, 24, 26, 27, 30, 32, 33, 38, 43, 45, 46, 47, 48, 49, 51, 52, 58, 59, 62, 66, 67, 68, 69,
     78, 79, 88, 89, 92, 93, 95, 96, 101, 104, 105, 107, 110, 111, 112, 113, 114, 115, 122, 126, 135, 136, 143},

    {1, 3, 5, 7, 9, 21, 22, 24, 25, 27, 30, 32, 33, 35, 36, 37, 39, 40, 44, 50, 52, 54, 58, 59, 65, 70, 71, 72, 73, 80,
     81, 82, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 95, 96, 101, 109, 133, 135, 136, 142, 143, 144},

    {2, 8, 9, 13, 14, 16, 17, 19, 24, 28, 30, 32, 34, 44, 45, 47, 48, 49, 51, 55, 58, 59, 60, 62, 76, 77, 84, 85, 86,
     87, 88, 92, 93, 94, 96, 101, 106, 108, 110, 111, 112, 113, 114, 115, 119, 127, 130, 133, 134, 143, 144},

    {14, 15, 16, 17, 24, 26, 28, 32, 33, 35, 36, 37, 43, 45, 46, 49, 51, 52, 54, 59, 62, 63, 64, 65, 66, 67, 68, 69, 84,
     85, 88, 89, 90, 91, 92, 93, 95, 96, 101, 110, 111, 122, 130, 135, 136, 139, 143},

    {11, 12, 14, 16, 17, 24, 26, 27, 29, 32, 33, 38, 43, 46, 47, 48, 49, 51, 53, 58, 59, 62, 65, 72, 73, 78, 79, 88, 89,
     92, 93, 95, 96, 101, 104, 105, 107, 112, 113, 114, 115, 126, 131, 135, 136, 143, 144},

    {14, 16, 17, 24, 26, 28, 32, 33, 35, 36, 37, 45, 49, 51, 52, 54, 59, 62, 63, 64, 65, 66, 67, 68, 69, 84, 85, 88, 90,
     91, 92, 93, 95, 96, 101, 110, 111, 135, 136, 140, 143},

]

# === 工具函数（使用配置中的范围） ===
def get_bounds():
    mins = np.array([EXPERIMENT_CONFIG['MIN_X'], EXPERIMENT_CONFIG['MIN_Y'], EXPERIMENT_CONFIG['MIN_Z']])
    maxs = np.array([EXPERIMENT_CONFIG['MAX_X'], EXPERIMENT_CONFIG['MAX_Y'], EXPERIMENT_CONFIG['MAX_Z']])
    return mins, maxs

def clip_state(state):
    mins, maxs = get_bounds()
    return np.clip(state, mins, maxs)

def denormalize_state(normalized_state):
    mins, maxs = get_bounds()
    return normalized_state * (maxs - mins) / 2 + (mins + maxs) / 2

def coverage_similarity(triggered, target_path):
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0
    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)

def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']
    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']
    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']
    return reward

# ==================== section8 触发函数（完整定义） ====================
def section8_hybrid_speed_torque_current(x, y, z):
    """第8类: 速度扭矩电流混合控制 (速度x, 扭矩y, 电流z) - 144个有效变异分支"""
    triggered = set()
    # 替换原self.standards的固定标准范围（基于代码逻辑推导合理值）
    speed_std = (1000, 2000)
    torque_std = (100, 200)
    current_std = (10, 30)

    # 删除的100%覆盖率分支: [1, 2, 3, 4, 5, 6] - 不包含在函数中
    # 删除的0%覆盖率分支: [39, 40, 59, 66, 74] - 不包含在函数中
    # 删除缺失分支: [137] (原编号148)

    # -------------------------- 1-32：速度扭矩电流基础检查（删除原1-6，保留7-38） --------------------------

    # 编号1 (原编号7)
    if [(x > 1400 and x < 1600 and y > 160 and y < 190)] != [(x > 1400 and x * 8 < 1600 and y > 160 and y < 190)]:
        triggered.add(1)

    # 编号2 (原编号8)
    if [(x > 1400 and x < 1600 and y > 160 and y < 190)] != [(x > 1400 and x < 1600 and y * 10 > 160 and y < 190)]:
        triggered.add(2)

    # 编号3 (原编号9)
    if [(x > 1350 and x < 1650 and z > 11 and z < 14)] != [(x > 1350 and x * 9 < 1650 and z > 11 and z < 14)]:
        triggered.add(3)

    # 编号4 (原编号10)
    if [(x > 1350 and x < 1650 and z > 11 and z < 14)] != [(x > 1350 and x < 1650 and z * 20 > 11 and z < 14)]:
        triggered.add(4)

    # 编号5 (原编号11)
    if [(y > 155 and y < 195 and z > 11 and z < 14)] != [(y > 155 and y * 7 < 195 and z > 11 and z < 14)]:
        triggered.add(5)

    # 编号6 (原编号12)
    if [(y > 155 and y < 195 and z > 11 and z < 14)] != [(y > 155 and y < 195 and z * 22 > 11 and z < 14)]:
        triggered.add(6)

    # 编号7 (原编号13)
    if [(x > 1480 and x < 1520)] != [(x > 1480 and x * 8 < 1520)]:
        triggered.add(7)

    # 编号8 (原编号14)
    if [(x > 1480 and x < 1520)] != [(x * 12 > 1480 and x < 1520)]:
        triggered.add(8)

    # 编号9 (原编号15)
    if [(y > 173 and y < 177)] != [(y * 12 > 173 and y < 177)]:
        triggered.add(9)

    # 编号10 (原编号16)
    if [(y > 173 and y < 177)] != [(y > 173 and y * 22 < 177)]:
        triggered.add(10)

    # 编号11 (原编号17)
    if [(z > 12.2 and z < 12.8)] != [(z * 78 > 12.2 and z < 12.8)]:
        triggered.add(11)

    # 编号12 (原编号18)
    if [(z > 12.2 and z < 12.8)] != [(66 > 12.2 and z < 12.8)]:
        triggered.add(12)

    # 编号13 (原编号19)
    if [(abs(x - 1500) < 40)] != [(abs(x - 1500) < 140)]:
        triggered.add(13)

    # 编号14 (原编号20)
    if [(abs(x - 1500) < 40)] != [(abs(x - 1500) < 990)]:
        triggered.add(14)

    # 编号15 (原编号21)
    if [(abs(y - 175) < 6)] != [(abs(y - 175) < 16)]:
        triggered.add(15)

    # 编号16 (原编号22)
    if [(abs(y - 175) < 6)] != [(abs(y - 175) < 226)]:
        triggered.add(16)

    # 编号17 (原编号23)
    if [(abs(z - 12.5) < 0.4)] != [(abs(z - 12.5) < 4)]:
        triggered.add(17)

    # 编号18 (原编号24)
    if [(abs(z - 12.5) < 0.4)] != [(abs(z - 12.5) < 1)]:
        triggered.add(18)

    # 编号19 (原编号25)
    if [(x > 1350 and y > 155 and z > 11)] != [(x > 1350 and y * 78 > 155 and z > 11)]:
        triggered.add(19)

    # 编号20 (原编号26)
    if [(x > 1350 and y > 155 and z > 11)] != [(x * 67 > 1350 and y > 155 and z > 11)]:
        triggered.add(20)

    # 编号21 (原编号27)
    if [(x < 1650 and y < 195 and z < 14)] != [(x < 1650 and y * 88 < 195 and z < 14)]:
        triggered.add(21)

    # 编号22 (原编号28)
    if [(x < 1650 and y < 195 and z < 14)] != [(x < 1650 and y < 195 and z * 8 < 14)]:
        triggered.add(22)

    # 编号23 (原编号29)
    if [((x / 10 + y + z * 10) > 350 and (x / 10 + y + z * 10) < 410)] != [
        ((x / 10 + y + z * 10) > 3150 and (x / 10 + y + z * 10) < 410)]:
        triggered.add(23)

    # 编号24 (原编号30)
    if [((x / 10 + y + z * 10) > 350 and (x / 10 + y + z * 10) < 410)] != [
        ((x / 10 + y + z * 10) > 350 and (x / 10 + y + z * 10) < 1410)]:
        triggered.add(24)

    # 编号25 (原编号31)
    if [(x * y / 1000 > 230 and x * y / 1000 < 310)] != [(x * y / 1000 > 1230 and x * y / 1000 < 310)]:
        triggered.add(25)

    # 编号26 (原编号32)
    if [(x * y / 1000 > 230 and x * y / 1000 < 310)] != [(x * y / 1000 > 230 and x * y / 1000 < 3110)]:
        triggered.add(26)

    # 编号27 (原编号33)
    if [(x * z / 100 > 160 and x * z / 100 < 220)] != [(x * z / 100 > 1610 and x * z / 100 < 220)]:
        triggered.add(27)

    # 编号28 (原编号34)
    if [(x * z / 100 > 160 and x * z / 100 < 220)] != [(x * z / 100 > 160 and x * z / 100 < 2120)]:
        triggered.add(28)

    # 编号29 (原编号35)
    if [(y * z > 2000 and y * z < 2600)] != [(y * z > 200 and y * z < 2600)]:
        triggered.add(29)

    # 编号30 (原编号36)
    if [(y * z > 2000 and y * z < 2600)] != [(y * z > 2000 and y * z < 260)]:
        triggered.add(30)

    # 编号31 (原编号37)
    if [((x / 10 + y + z * 10) / 3 > 116 and (x / 10 + y + z * 10) / 3 < 136)] != [
        ((x / 10 + y + z * 10) / 5 > 116 and (x / 10 + y + z * 10) / 3 < 136)]:
        triggered.add(31)

    # 编号32 (原编号38)
    if [((x / 10 + y + z * 10) / 3 > 116 and (x / 10 + y + z * 10) / 3 < 136)] != [
        ((x / 10 + y + z * 10) / 3 > 116 and (x / 10 + y + z * 10) / 11 < 136)]:
        triggered.add(32)

    # 跳过原编号39, 40 (0%覆盖率分支)

    # -------------------------- 33-58：驱动系统分析（原41-80，删除原39、40） --------------------------

    # 编号33 (原编号41)
    if [(x / (y * 10) > 0.8 and x / (y * 10) < 1.0)] != [(x / (y * 10) > 8 and x / (y * 10) < 1.0)]:
        triggered.add(33)

    # 编号34 (原编号42)
    if [(x / (y * 10) > 0.8 and x / (y * 10) < 1.0)] != [(x / (y * 10) > 0.8 and x / (y * 10) < 10)]:
        triggered.add(34)

    # 编号35 (原编号43)
    if [(x / z > 110 and x / z < 140)] != [(x / z > 1220 and x / z < 140)]:
        triggered.add(35)

    # 编号36 (原编号44)
    if [(x / z > 110 and x / z < 140)] != [(x / z > 110 and x / z < 14)]:
        triggered.add(36)

    # 编号37 (原编号45)
    if [(y / z > 12 and y / z < 16)] != [(y / z > 112 and y / z < 16)]:
        triggered.add(37)

    # 编号38 (原编号46)
    if [(y / z > 12 and y / z < 16)] != [(y / z > 12 and y / z < 1226)]:
        triggered.add(38)

    # 编号39 (原编号47)
    if [((x - 1200) / 10 + (y - 150) > 35 and (x - 1200) / 10 + (y - 150) < 55)] != [
        ((x - 1200) / 10 + (y - 150) > 315 and (x - 1200) / 10 + (y - 150) < 55)]:
        triggered.add(39)

    # 编号40 (原编号48)
    if [((x - 1200) / 10 + (y - 150) > 35 and (x - 1200) / 10 + (y - 150) < 55)] != [
        ((x - 1200) / 10 + (y - 220) > 35 and (x - 1200) / 10 + (y - 150) < 55)]:
        triggered.add(40)

    # 编号41 (原编号49)
    if [((x - 1200) / 10 + (z - 10) * 5 > 40 and (x - 1200) / 10 + (z - 10) * 5 < 60)] != [
        ((x - 1200) / 10 + (z - 10) * 15 > 40 and (x - 1200) / 10 + (z - 10) * 5 < 60)]:
        triggered.add(41)

    # 编号42 (原编号50)
    if [((x - 1200) / 10 + (z - 10) * 5 > 40 and (x - 1200) / 10 + (z - 10) * 5 < 60)] != [
        ((x - 1200) / 10 + (z - 10) * 115 > 40 and (x - 1200) / 10 + (z - 10) * 5 < 60)]:
        triggered.add(42)

    # 编号43 (原编号51)
    if [((y - 150) + (z - 10) * 2 > 30 and (y - 150) + (z - 10) * 2 < 50)] != [
        ((y - 150) + (z - 10) * 2 > 302 and (y - 150) + (z - 10) * 2 < 50)]:
        triggered.add(43)

    # 编号44 (原编号52)
    if [((y - 150) + (z - 10) * 2 > 30 and (y - 150) + (z - 10) * 2 < 50)] != [
        ((y - 150) + (z - 10) * 211 > 30 and (y - 150) + (z - 10) * 2 < 50)]:
        triggered.add(44)

    # 编号45 (原编号53)
    if [(abs((x - 1500) / 10 - (y - 175)) < 12)] != [(abs((x - 1500) / 10 - (y - 175)) < 112)]:
        triggered.add(45)

    # 编号46 (原编号54)
    if [(abs((x - 1500) / 10 - (y - 175)) < 12)] != [(abs((x - 1500) / 101 - (y - 175)) < 12)]:
        triggered.add(46)

    # 编号47 (原编号55)
    if [(abs((x - 1500) / 100 - (z - 12.5)) < 3)] != [(abs((x - 1500) / 100 - (z - 12.5)) < 43)]:
        triggered.add(47)

    # 编号48 (原编号56)
    if [(abs((x - 1500) / 100 - (z - 12.5)) < 3)] != [(abs((x - 1500) / 100 - (z - 12.5)) < 13)]:
        triggered.add(48)

    # 编号49 (原编号57)
    if [(abs((y - 175) - (z - 12.5) * 12) < 8)] != [(abs((y - 175) - (z - 12.5) * 12) < 118)]:
        triggered.add(49)

    # 编号50 (原编号58)
    if [(abs((y - 175) - (z - 12.5) * 12) < 8)] != [(abs((y - 175) - (z - 12.5) * 232) < 8)]:
        triggered.add(50)

    # 跳过原编号59 (0%覆盖率分支)

    # 编号51 (原编号60)
    if [(x / (y * 10 + 500) > 0.7 and x / (y * 10 + 500) < 0.9)] != [
        (x / (y * 10 + 500) > 337 and x / (y * 10 + 500) < 0.9)]:
        triggered.add(51)

    # 编号52 (原编号61)
    if [(y / (z + 5) > 9 and y / (z + 5) < 13)] != [(y / (z + 5) > 119 and y / (z + 5) < 13)]:
        triggered.add(52)

    # 编号53 (原编号62)
    if [(y / (z + 5) > 9 and y / (z + 5) < 13)] != [(y / (z + 5) > 9 and y / (z + 5) < 123)]:
        triggered.add(53)

    # 编号54 (原编号63)
    if [(z / (x / 150) > 1.1 and z / (x / 150) < 1.5)] != [(z / (x / 150) > 11 and z / (x / 150) < 1.5)]:
        triggered.add(54)

    # 编号55 (原编号64)
    if [(z / (x / 150) > 1.1 and z / (x / 150) < 1.5)] != [(z / (x / 150) > 1.1 and z / (x / 150) < 15)]:
        triggered.add(55)

    # 编号56 (原编号65)
    if [((x * y * z / 10000) > 28 and (x * y * z / 10000) < 38)] != [
        ((x * y * z / 10000) > 28 and (x * y * z / 10000) < 238)]:
        triggered.add(56)

    # 跳过原编号66 (0%覆盖率分支)

    # 编号57 (原编号67)
    if [(x / 20 + y * 0.6 + z * 8 > 260 and x / 20 + y * 0.6 + z * 8 < 300)] != [
        (x / 20 + y * 0.6 + z * 10 > 260 and x / 20 + y * 0.6 + z * 8 < 300)]:
        triggered.add(57)

    # 编号58 (原编号68)
    if [(x / 20 + y * 0.6 + z * 8 > 260 and x / 20 + y * 0.6 + z * 8 < 300)] != [
        (x / 20 + y * 0.6 + z * 8 > 260 and x / 20 + y * 6 + z * 8 < 300)]:
        triggered.add(58)

    # 编号59 (原编号69)
    if [((x / 1500) ** 0.4 * (y / 175) ** 0.4 * (z / 12.5) ** 0.2 > 0.93)] != [
        ((x / 1500) ** 0.4 * (y / 175) ** 0.4 * (z / 12.5) ** 0.2 > 93)]:
        triggered.add(59)

    # 编号60 (原编号70)
    if [((x / 1500) ** 0.4 * (y / 175) ** 0.4 * (z / 12.5) ** 0.2 > 0.93)] != [
        ((x / 1500) ** 0.4 * (y / 175) ** 4 * (z / 12.5) ** 0.2 > 0.93)]:
        triggered.add(60)

    # 编号61 (原编号71)
    if [((x - 1500) / 10 * (y - 175) > -200 and (x - 1500) / 10 * (y - 175) < 200)] != [
        ((x - 1500) / 110 * (y - 175) > -200 and (x - 1500) / 10 * (y - 175) < 200)]:
        triggered.add(61)

    # 编号62 (原编号72)
    if [((x - 1500) / 10 * (y - 175) > -200 and (x - 1500) / 10 * (y - 175) < 200)] != [
        ((x - 1500) / 10 * (y - 175) > -200 and (x - 1500) / 110 * (y - 175) < 200)]:
        triggered.add(62)

    # 编号63 (原编号73)
    if [((x - 1500) / 100 * (z - 12.5) > -20 and (x - 1500) / 100 * (z - 12.5) < 20)] != [
        ((x - 1500) / 100 * (z - 12.5) > -20 and (x - 1500) / 10 * (z - 12.5) < 20)]:
        triggered.add(63)

    # 跳过原编号74 (0%覆盖率分支)

    # 编号64 (原编号75)
    if [((y - 175) * (z - 12.5) > -20 and (y - 175) * (z - 12.5) < 20)] != [
        ((y - 175) * (z - 12.5) > -20 and (y - 175) * (z - 12.5) < 210)]:
        triggered.add(64)

    # 编号65 (原编号76)
    if [((y - 175) * (z - 12.5) > -20 and (y - 175) * (z - 12.5) < 20)] != [((y - 175) * (z - 125) < 20)]:
        triggered.add(65)

    # 编号66 (原编号77)
    if [(abs(x / 1500 + y / 175 + z / 12.5 - 3) < 0.2)] != [(abs(x / 1500 + y / 175 + z / 12.5 - 3) < 112)]:
        triggered.add(66)

    # 编号67 (原编号78)
    if [(abs(x / 1500 + y / 175 + z / 12.5 - 3) < 0.2)] != [(abs(x / 1500 + y / 175 + z / 12.5 - 3) < 2222)]:
        triggered.add(67)

    # 编号68 (原编号79)
    if [(abs((x / 1500) * (y / 175) * (z / 12.5) - 1) < 0.15)] != [
        (abs((x / 1500) * (y / 175) * (z / 12.5) - 1) < 115)]:
        triggered.add(68)

    # 编号69 (原编号80)
    if [(abs((x / 1500) * (y / 175) * (z / 12.5) - 1) < 0.15)] != [
        (abs((x / 1500) * (y / 175) * (z / 12.5) - 1) < 2225)]:
        triggered.add(69)

    # -------------------------- 70-107：动力传动协调（原81-120） --------------------------

    # 编号70 (原编号81)
    if [(x > 1480 and x < 1520 and y > 172 and y < 178 and z > 12.2 and z < 12.8)] != [
        (x > 1480 and x < 1520 and y > 172 and y < 178 and z > 12.2 and 8 < 12.8)]:
        triggered.add(70)

    # 编号71 (原编号82)
    if [(x > 1480 and x < 1520 and y > 172 and y < 178 and z > 12.2 and z < 12.8)] != [
        (x > 1480 and x < 1520 and y > 172 and y < 178 and z > 12.2 and 10 < 12.8)]:
        triggered.add(71)

    # 编号72 (原编号83)
    if [(abs(x / (y * 10) - 0.86) < 0.06 and abs(z - 12.5) < 0.3)] != [
        (abs(x / (y * 10) - 0.86) < 0.06 and abs(z - 12.5) < 3)]:
        triggered.add(72)

    # 编号73 (原编号84)
    if [(abs(x / (y * 10) - 0.86) < 0.06 and abs(z - 12.5) < 0.3)] != [
        (abs(x / (y * 10) - 0.86) < 0.06 and abs(z - 12.5) < 13)]:
        triggered.add(73)

    # 编号74 (原编号85)
    if [(x * y / 1000 > 258 and x * y / 1000 < 268 and z > 12.3 and z < 12.7)] != [
        (x * y / 1000 > 258 and x * y / 1000 < 268 and 8 * z > 12.3 and z < 12.7)]:
        triggered.add(74)

    # 编号75 (原编号86)
    if [(x * y / 1000 > 258 and x * y / 1000 < 268 and z > 12.3 and z < 12.7)] != [
        (x * y / 1000 > 258 and x * y / 1000 < 268 and z * 99 > 12.3 and z < 12.7)]:
        triggered.add(75)

    # 编号76 (原编号87)
    if [((x / 10 + y) / 2 > 137 and (x / 10 + y) / 2 < 143 and z > 12.2 and z < 12.8)] != [
        ((x / 10 + y) / 2 > 137 and (x / 10 + y) / 2 < 143 and z > 12.2 and 8 < 12.8)]:
        triggered.add(76)

    # 编号77 (原编号88)
    if [((x / 10 + y) / 2 > 137 and (x / 10 + y) / 2 < 143 and z > 12.2 and z < 12.8)] != [
        ((x / 10 + y) / 2 > 137 and (x / 10 + y) / 2 < 143 and z > 12.2 and 10 < 12.8)]:
        triggered.add(77)

    # 编号78 (原编号89)
    if [(abs(x / 10 - y) < 25 and z > 12.2 and z < 12.8)] != [(abs(x / 10 - y) < 25 and z * 12 > 12.2 and z < 12.8)]:
        triggered.add(78)

    # 编号79 (原编号90)
    if [(abs(x / 10 - y) < 25 and z > 12.2 and z < 12.8)] != [(abs(x / 10 - y) < 25 and z * 89 > 12.2 and z < 12.8)]:
        triggered.add(79)

    # 编号80 (原编号91)
    if [(math.sqrt((x / 10 - 150) ** 2 + (y - 175) ** 2) < 6 and abs(z - 12.5) < 0.3)] != [
        (math.sqrt((x / 10 - 150) ** 2 + (y - 175) ** 2) < 6 and abs(z - 12.5) < 3)]:
        triggered.add(80)

    # 编号81 (原编号92)
    if [(math.sqrt((x / 10 - 150) ** 2 + (y - 175) ** 2) < 6 and abs(z - 12.5) < 0.3)] != [
        (math.sqrt((x / 10 - 150) ** 2 + (y - 175) ** 2) < 6 and abs(z - 12.5) < 13)]:
        triggered.add(81)

    # 编号82 (原编号93)
    if [(x / 1500 > 0.98 and x / 1500 < 1.02 and y / 175 > 0.97 and y / 175 < 1.03)] != [
        (x / 1500 > 0.98 and x / 1500 < 1.02 and y / 175 > 97 and y / 175 < 1.03)]:
        triggered.add(82)

    # 编号83 (原编号94)
    if [(x / 1500 > 0.98 and x / 1500 < 1.02 and y / 175 > 0.97 and y / 175 < 1.03)] != [
        (x / 1500 > 0.98 and x / 1500 < 1.02 and y * 8 / 175 > 0.97 and y / 175 < 1.03)]:
        triggered.add(83)

    # 编号84 (原编号95)
    if [(z / 12.5 > 0.976 and z / 12.5 < 1.024)] != [(z / 12.5 > 0.976 and z / 125 < 1.024)]:
        triggered.add(84)

    # 编号85 (原编号96)
    if [(z / 12.5 > 0.976 and z / 12.5 < 1.024)] != [(z / 12.5 > 0.976 and z / 200 < 1.024)]:
        triggered.add(85)

    # 编号86 (原编号97)
    if [((x / 1500 + y / 175 + z / 12.5) / 3 > 0.98 and (x / 1500 + y / 175 + z / 12.5) / 3 < 1.02)] != [
        ((x / 1500 + y / 175 + z / 12.5) / 3 > 98 and (x / 1500 + y / 175 + z / 12.5) / 3 < 1.02)]:
        triggered.add(86)

    # 编号87 (原编号98)
    if [((x / 1500 + y / 175 + z / 12.5) / 3 > 0.98 and (x / 1500 + y / 175 + z / 12.5) / 3 < 1.02)] != [
        ((x / 1500 + y / 175 + z / 12.5) / 23 > 0.98 and (x / 1500 + y / 175 + z / 12.5) / 3 < 1.02)]:
        triggered.add(87)

    # 编号88 (原编号99)
    if [(max(abs(x / 1500 - 1), abs(y / 175 - 1), abs(z / 12.5 - 1)) < 0.025)] != [
        (max(abs(x / 1500 - 1), abs(y / 175 - 1), abs(z / 12.5 - 1)) < 0.55)]:
        triggered.add(88)

    # 编号89 (原编号100)
    if [(max(abs(x / 1500 - 1), abs(y / 175 - 1), abs(z / 12.5 - 1)) < 0.025)] != [
        (max(abs(x / 1500 - 1), abs(y / 175 - 1), abs(z / 12.5 - 1)) < 0.25)]:
        triggered.add(89)

    # 编号90 (原编号101)
    if [(min(x / 1500, y / 175, z / 12.5) > 0.98)] != [(min(x / 1500, y / 175, z / 12.5) > 498)]:
        triggered.add(90)

    # 编号91 (原编号102)
    if [(min(x / 1500, y / 175, z / 12.5) > 0.98)] != [(min(x / 1500, y / 1375, z / 12.5) > 0.98)]:
        triggered.add(91)

    # 编号92 (原编号103)
    if [(max(x / 1500, y / 175, z / 12.5) < 1.02)] != [(max(x / 1500, y / 175, z / 12.5) < 302)]:
        triggered.add(92)

    # 编号93 (原编号104)
    if [(max(x / 1500, y / 175, z / 12.5) < 1.02)] != [(max(x / 1500, y / 175, z / 12.5) < 102)]:
        triggered.add(93)

    # 编号94 (原编号105)
    if [(abs(max(x / 10, y, z * 10) - min(x / 10, y, z * 10)) < 30)] != [
        (abs(max(x / 10, y, z * 120) - min(x / 10, y, z * 10)) < 30)]:
        triggered.add(94)

    # 编号95 (原编号106)
    if [(abs(max(x / 10, y, z * 10) - min(x / 10, y, z * 10)) < 30)] != [
        (abs(max(x / 10, y, z * 10) - min(x / 10, y, z * 10)) < 230)]:
        triggered.add(95)

    # 编号96 (原编号107)
    if [((x / 10 + y + z * 10) > 378 and (x / 10 + y + z * 10) < 382)] != [
        ((x / 10 + y + z * 10) > 378 and (x / 10 + y + z * 10) < 3112)]:
        triggered.add(96)

    # 编号97 (原编号108)
    if [((x / 10 + y + z * 10) > 378 and (x / 10 + y + z * 10) < 382)] != [
        ((x / 10 + y + z * 10) > 3781 and (x / 10 + y + z * 10) < 382)]:
        triggered.add(97)

    # 编号98 (原编号109)
    if [((x * y * z / 10000) > 32.5 and (x * y * z / 10000) < 33.5)] != [
        ((x * y * z / 10000) > 32.5 and (x * y * 2 / 10000) < 33.5)]:
        triggered.add(98)

    # 编号99 (原编号110)
    if [((x * y * z / 10000) > 32.5 and (x * y * z / 10000) < 33.5)] != [
        ((x * y * z / 10000) > 32.5 and (x * y * 2.4 / 10000) < 33.5)]:
        triggered.add(99)

    # 编号100 (原编号111)
    if [(abs((x / 10 + y + z * 10) / 3 - 126.7) < 1.5)] != [(abs((x / 10 + y + z * 10) / 3 - 126.7) < 15)]:
        triggered.add(100)

    # 编号101 (原编号112)
    if [(abs((x / 10 + y + z * 10) / 3 - 126.7) < 1.5)] != [(abs((x / 10 + y + z * 10) / 3 - 126.7) < 115)]:
        triggered.add(101)

    # 编号102 (原编号113)
    if [(x > 1495 and x < 1505 and y > 174 and y < 176)] != [(x > 1495 and x < 1505 and y * 10 > 174 and y < 176)]:
        triggered.add(102)

    # 编号103 (原编号114)
    if [(x > 1495 and x < 1505 and y > 174 and y < 176)] != [(x > 1495 and x < 1505 and y * 88 > 174 and y < 176)]:
        triggered.add(103)

    # 编号104 (原编号115)
    if [(z > 12.45 and z < 12.55)] != [(z * 10 > 12.45 and z < 12.55)]:
        triggered.add(104)

    # 编号105 (原编号116)
    if [(z > 12.45 and z < 12.55)] != [(z * 8 > 12.45 and z < 12.55)]:
        triggered.add(105)

    # -------------------------- 106-144：动力控制优化（原117-156，删除原148） --------------------------

    # 编号106 (原编号117)
    if [(x < 1350 or y < 160 or z < 11.5)] != [(x < 1350 or y * 8 < 160 or z < 11.5)]:
        triggered.add(106)

    # 编号107 (原编号118)
    if [(x < 1350 or y < 160 or z < 11.5)] != [(x < 1350 or y < 160 or z * 89 < 11.5)]:
        triggered.add(107)

    # 编号108 (原编号119)
    if [(x > 1650 or y > 190 or z > 13.5)] != [(x > 1650 or y > 190 or z > 135)]:
        triggered.add(108)

    # 编号109 (原编号120)
    if [(x > 1650 or y > 190 or z > 13.5)] != [(x > 1650 or y * 87 > 190 or z > 13.5)]:
        triggered.add(109)

    # 编号110 (原编号121)
    if [(abs(x / (y * 10) - 0.86) > 0.1)] != [(abs(x / (y * 10) - 0.86) > 111)]:
        triggered.add(110)

    # 编号111 (原编号122)
    if [(abs(x / (y * 10) - 0.86) > 0.1)] != [(abs(x / (y * 10) - 0.86) > 1671)]:
        triggered.add(111)

    # 编号112 (原编号123)
    if [(abs(x / z - 120) > 15)] != [(abs(x / z - 120) > 1522)]:
        triggered.add(112)

    # 编号113 (原编号124)
    if [(abs(x / z - 120) > 15)] != [(abs(x / z - 120) > 115)]:
        triggered.add(113)

    # 编号114 (原编号125)
    if [(abs(y / z - 14) > 2)] != [(abs(y / z - 14) > 112)]:
        triggered.add(114)

    # 编号115 (原编号126)
    if [(abs(y / z - 14) > 2)] != [(abs(y / z - 14) > 211)]:
        triggered.add(115)

    # 编号116 (原编号127)
    if [((x / 10 + y + z * 10) < 360 or (x / 10 + y + z * 10) > 400)] != [
        ((x / 10 + y + z * 10) < 36 or (x / 10 + y + z * 10) > 400)]:
        triggered.add(116)

    # 编号117 (原编号128)
    if [((x / 10 + y + z * 10) < 360 or (x / 10 + y + z * 10) > 400)] != [
        ((x / 10 + y + z * 10) < 989 or (x / 10 + y + z * 10) > 400)]:
        triggered.add(117)

    # 编号118 (原编号129)
    if [((x * y * z / 10000) < 30 or (x * y * z / 10000) > 36)] != [
        ((x * y * z / 10000) < 30 or (x * y * z / 10000) > 136)]:
        triggered.add(118)

    # 编号119 (原编号130)
    if [((x * y * z / 10000) < 30 or (x * y * z / 10000) > 36)] != [
        ((x * y * z / 10000) < 30 or (x * y * z / 10000) > 316)]:
        triggered.add(119)

    # 编号120 (原编号131)
    if [(x < 1300 and y < 160)] != [(x < 1300 and y * 878 < 160)]:
        triggered.add(120)

    # 编号121 (原编号132)
    if [(x < 1300 and y < 160)] != [(x * 71 < 1300 and y < 160)]:
        triggered.add(121)

    # 编号122 (原编号133)
    if [(x > 1700 and y > 190)] != [(x > 1700 and y * 78 > 190)]:
        triggered.add(122)

    # 编号123 (原编号134)
    if [(x > 1700 and y > 190)] != [(x * 78 > 1700 and y > 190)]:
        triggered.add(123)

    # 编号124 (原编号135)
    if [(x < 1300 and z < 11)] != [(x < 1300 and z * 91 < 11)]:
        triggered.add(124)

    # 编号125 (原编号136)
    if [(x < 1300 and z < 11)] != [(x * 12 < 1300 and z < 11)]:
        triggered.add(125)

    # 编号126 (原编号137)
    if [(x > 1700 and z > 14)] != [(x > 1700 and z * 21 > 14)]:
        triggered.add(126)

    # 编号127 (原编号138)
    if [(x > 1700 and z > 14)] != [(x * 123 > 1700 and z > 14)]:
        triggered.add(127)

    # 编号128 (原编号139)
    if [(y < 160 and z < 11)] != [(y < 160 and z * 78 < 11)]:
        triggered.add(128)

    # 编号129 (原编号140)
    if [(y < 160 and z < 11)] != [(y * 8 < 160 and z < 11)]:
        triggered.add(129)

    # 编号130 (原编号141)
    if [(y > 190 and z > 14)] != [(y * 8 > 190 and z > 14)]:
        triggered.add(130)

    # 编号131 (原编号142)
    if [(y > 190 and z > 14)] != [(y > 190 and z * 9 > 14)]:
        triggered.add(131)

    # 编号132 (原编号143)
    if [(x < 1250 or x > 1750)] != [(x * 67 < 1250 or x > 1750)]:
        triggered.add(132)

    # 编号133 (原编号144)
    if [(x < 1250 or x > 1750)] != [(x < 1250 or x * 53 > 1750)]:
        triggered.add(133)

    # 编号134 (原编号145)
    if [(y < 145 or y > 205)] != [(y * 67 < 145 or y > 205)]:
        triggered.add(134)

    # 编号135 (原编号146)
    if [(y < 145 or y > 205)] != [(y < 145 or y * 67 > 205)]:
        triggered.add(135)

    # 编号136 (原编号147)
    if [(z < 9.5 or z > 15.5)] != [(z < 9.5 or z * 67 > 15.5)]:
        triggered.add(136)

    # 跳过原编号148 (缺失分支)

    # 编号137 (原编号149)
    if [(x < 1200 and y < 155 and z < 10.5)] != [(x * 67 < 1200 and y < 155 and z < 10.5)]:
        triggered.add(137)

    # 编号138 (原编号150)
    if [(x < 1200 and y < 155 and z < 10.5)] != [(x < 1200 and y < 55 and z < 10.5)]:
        triggered.add(138)

    # 编号139 (原编号151)
    if [(x > 1800 and y > 195 and z > 14.5)] != [(x > 1800 and y * 78 > 195 and z > 14.5)]:
        triggered.add(139)

    # 编号140 (原编号152)
    if [(x > 1800 and y > 195 and z > 14.5)] != [(x > 1800 and y > 966 and z > 14.5)]:
        triggered.add(140)

    # 编号141 (原编号153)
    if [((x / 10 - 150) ** 2 + (y - 175) ** 2 + (z - 12.5) ** 2 * 100) > 150] != [
        ((x / 10 - 150) ** 6 + (y - 175) ** 2 + (z - 12.5) ** 2 * 100) > 150]:
        triggered.add(141)

    # 编号142 (原编号154)
    if [((x / 10 - 150) ** 2 + (y - 175) ** 2 + (z - 12.5) ** 2 * 100) > 150] != [
        ((x / 10 - 150) ** 2 + (y - 175) ** 8 + (z - 12.5) ** 2 * 100) > 150]:
        triggered.add(142)

    # 编号143 (原编号155)
    if [(abs((x * y * z / 10000) / 33 - 1) > 0.18)] != [(abs((x * y * z / 10000) / 33 - 1) > 18)]:
        triggered.add(143)

    # 编号144 (原编号156)
    if [(abs((x * y * z / 10000) / 33 - 1) > 0.18)] != [(abs((x * y * z / 10000) / 33 - 1) > 10)]:
        triggered.add(144)

    return triggered

# ========== 设置执行函数为 section8 ==========
execute_Tr = section8_hybrid_speed_torque_current

# === DQN 网络 ===
class DQNNetwork(nn.Module):
    def __init__(self, action_size=30):
        super(DQNNetwork, self).__init__()
        hidden_dim = EXPERIMENT_CONFIG['HIDDEN_DIM']
        self.fc1 = nn.Linear(3, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)
        self.output = nn.Linear(hidden_dim, action_size)

    def forward(self, x):
        x = torch.relu(self.fc1(x))
        x = torch.relu(self.fc2(x))
        x = torch.relu(self.fc3(x))
        return self.output(x)

# === PathReplayBuffer ===
class PathReplayBuffer:
    def __init__(self, path_idx, capacity=20000):
        self.path_idx = path_idx
        self.capacity = capacity
        self.buffer = deque(maxlen=capacity)
        self.similarities = deque(maxlen=capacity)

    def push(self, state, action, reward, next_state, done, similarity):
        self.buffer.append((state, action, reward, next_state, done))
        self.similarities.append(similarity)

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return None
        batch = random.sample(self.buffer, batch_size)
        state, action, reward, next_state, done = map(np.stack, zip(*batch))
        return (
            torch.FloatTensor(state).to(device),
            torch.LongTensor(action).to(device),
            torch.FloatTensor(reward).to(device),
            torch.FloatTensor(next_state).to(device),
            torch.BoolTensor(done).to(device)
        )

    def get_top_k(self, k=20):
        if len(self.buffer) == 0:
            return []
        samples_with_sim = list(zip(self.buffer, self.similarities))
        samples_with_sim.sort(key=lambda x: x[1], reverse=True)
        top_k = samples_with_sim[:k]
        results = []
        for (state, _, _, _, _), similarity in top_k:
            original_state = denormalize_state(state)
            original_state_int = np.round(original_state).astype(int)
            results.append({
                'state': original_state_int,
                'similarity': similarity,
                'triggered': execute_Tr(*original_state_int)
            })
        return results

    def __len__(self):
        return len(self.buffer)

# === ImprovedDQNAgent ===
class ImprovedDQNAgent:
    def __init__(self, num_paths, action_size=30):
        self.action_size = action_size
        self.num_paths = num_paths
        self.epsilon = 0.9
        self.epsilon_min = 0.1
        self.epsilon_decay = 0.995

        self.q_network = DQNNetwork(action_size).to(device)
        self.target_network = DQNNetwork(action_size).to(device)
        self.optimizer = optim.Adam(self.q_network.parameters(), lr=EXPERIMENT_CONFIG['LEARNING_RATE'])

        self.replay_buffers = {}
        for path_idx in range(num_paths):
            self.replay_buffers[path_idx] = PathReplayBuffer(path_idx, EXPERIMENT_CONFIG['REPLAY_BUFFER_CAPACITY'])

        self.replay_train_count = 0
        self.update_target_network()

    def discrete_to_action_delta(self, action_idx):
        delta_values = [5, 3, 2, 1, 0.5, -0.5, -1, -2, -3, -5]
        if action_idx >= 30:
            action_idx = action_idx % 30
        dim = action_idx // 10
        delta_idx = action_idx % 10
        delta = delta_values[delta_idx]
        action_delta = np.zeros(3)
        action_delta[dim] = delta
        return action_delta

    def get_action(self, state):
        mins, maxs = get_bounds()
        normalized_state = (state - (mins + maxs) / 2) / ((maxs - mins) / 2)
        if random.random() < self.epsilon:
            action_idx = random.randint(0, self.action_size - 1)
        else:
            state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)
            with torch.no_grad():
                q_values = self.q_network(state_tensor)
                action_idx = q_values.argmax().item()
        action_delta = self.discrete_to_action_delta(action_idx)
        return action_delta, action_idx

    def store_experience(self, path_idx, state, action_idx, reward, next_state, done, similarity):
        mins, maxs = get_bounds()
        normalized_state = (state - (mins + maxs) / 2) / ((maxs - mins) / 2)
        normalized_next_state = (next_state - (mins + maxs) / 2) / ((maxs - mins) / 2)
        self.replay_buffers[path_idx].push(
            normalized_state, action_idx, reward,
            normalized_next_state, done, similarity
        )

    def replay_train(self, path_idx):
        batch = self.replay_buffers[path_idx].sample(EXPERIMENT_CONFIG['REPLAY_BATCH_SIZE'])
        if batch is None:
            return
        states, actions, rewards, next_states, dones = batch
        current_q_values = self.q_network(states).gather(1, actions.unsqueeze(1))
        with torch.no_grad():
            next_q_values = self.target_network(next_states).max(1)[0]
            target_q_values = rewards + (0.99 * next_q_values * ~dones)
        loss = F.mse_loss(current_q_values.squeeze(), target_q_values)
        self.optimizer.zero_grad()
        loss.backward()
        torch.nn.utils.clip_grad_norm_(self.q_network.parameters(), 1.0)
        self.optimizer.step()
        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay
        self.replay_train_count += 1
        if self.replay_train_count % 2 == 0:
            self.update_target_network()

    def update_target_network(self):
        self.target_network.load_state_dict(self.q_network.state_dict())

    def get_all_top_k(self, k=20):
        results = {}
        for path_idx in range(self.num_paths):
            results[path_idx] = self.replay_buffers[path_idx].get_top_k(k)
        return results

    def get_buffer_stats(self):
        stats = {}
        for path_idx in range(self.num_paths):
            stats[path_idx] = len(self.replay_buffers[path_idx])
        return stats

# === 性能计算 ===
def calculate_run_performance(run_idx, dqn_results, training_time, total_steps, update_count, agent):
    target_paths = targetPaths
    num_paths = len(target_paths)
    all_similarities = []
    total_reward = 0
    total_samples = 0
    all_rewards = []
    for path_idx in range(num_paths):
        samples = dqn_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']
            total_reward += reward
            all_rewards.append(reward)
            all_similarities.append(similarity)
            total_samples += 1
    average_reward = total_reward / total_samples if total_samples > 0 else 0
    convergence = np.mean(all_similarities) if all_similarities else 0
    environment_adaptability = 1 / (np.std(all_similarities) + 1e-8) if len(all_similarities) > 1 else 0
    generalization_ability = convergence
    computational_efficiency = total_steps / training_time if training_time > 0 else 0
    policy_update_frequency = update_count / training_time if training_time > 0 else 0
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0
    return {
        'Run': run_idx + 1,
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }

# === Excel导出 ===
def export_to_excel(all_dqn_results, all_performance_data, target_paths, output_path="DQN_20_run.xlsx"):
    print("\n导出Excel...")
    all_dqn_summary_data = []
    all_dqn_detailed_data = []
    for run_idx, (dqn_results, performance_data) in enumerate(zip(all_dqn_results, all_performance_data)):
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = dqn_results[path_idx]
            if len(samples) == 0:
                all_dqn_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue
            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = 'Yes' if perfect_count > 0 else 'No'
            all_dqn_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })
            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']
                all_dqn_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Sample ID': sample_idx + 1,
                    'X': int(state[0]),
                    'Y': int(state[1]),
                    'Z': int(state[2]),
                    'Similarity': round(similarity, 4),
                    'Perfect Coverage': 'Yes' if abs(similarity - 1.0) < 0.001 else 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path))),
                    'Triggered Rules': ', '.join(map(str, sorted(triggered))),
                    'Hit Rule Count': len(target_path.intersection(triggered)),
                    'Target Rule Count': len(target_path)
                })
    dqn_summary_df = pd.DataFrame(all_dqn_summary_data)
    dqn_detailed_df = pd.DataFrame(all_dqn_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        dqn_summary_df.to_excel(writer, sheet_name='DQNPath', index=False)
        dqn_detailed_df.to_excel(writer, sheet_name='DQNDetailed Sample Data', index=False)
        selected_columns = [
            'Run', 'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df[selected_columns].to_excel(writer, sheet_name='Metric', index=False)
        workbook = writer.book
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        for sheet_name in ['DQNPath', 'DQNDetailed Sample Data', 'Metric']:
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
    print(f"Excel已保存: {output_path}")

# === 训练流程 ===
def train_dqn_workflow():
    print("=" * 80)
    print("开始DQN训练")
    print(f"路径数: {len(targetPaths)}, 每路径样本数: {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']}")
    print("=" * 80)
    num_paths = len(targetPaths)
    agent = ImprovedDQNAgent(num_paths=num_paths)
    start_time = time.time()
    total_steps = 0

    # 生成初始样本
    path_samples = {}
    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            state = np.array([
                random.randint(EXPERIMENT_CONFIG['MIN_X'], EXPERIMENT_CONFIG['MAX_X']),
                random.randint(EXPERIMENT_CONFIG['MIN_Y'], EXPERIMENT_CONFIG['MAX_Y']),
                random.randint(EXPERIMENT_CONFIG['MIN_Z'], EXPERIMENT_CONFIG['MAX_Z'])
            ], dtype=np.float32)
            samples.append(state)
        path_samples[path_idx] = samples

    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size
    num_rounds = EXPERIMENT_CONFIG['NUM_ROUNDS']

    for path_idx in range(num_paths):
        target_path = targetPaths[path_idx]
        print(f"\n开始训练路径 {path_idx+1}/{num_paths}, 目标规则数: {len(target_path)}")
        for round_idx in range(num_rounds):
            print(f"  轮次 {round_idx+1}/{num_rounds}")
            for batch_idx in range(num_batches):
                batch_samples = path_samples[path_idx][batch_idx*batch_size:(batch_idx+1)*batch_size]
                batch_rewards = []
                batch_similarities = []
                for initial_state in batch_samples:
                    state = initial_state.copy()
                    episode_reward = 0
                    final_similarity = 0
                    for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                        action_delta, action_idx = agent.get_action(state)
                        next_state = state + action_delta
                        next_state = clip_state(next_state)
                        triggered = execute_Tr(*next_state)
                        reward = unified_reward_function(triggered, target_path)
                        similarity = coverage_similarity(triggered, target_path)
                        done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)
                        agent.store_experience(path_idx, state, action_idx, reward, next_state, done, similarity)
                        state = next_state
                        episode_reward += reward
                        final_similarity = similarity
                        total_steps += 1
                    batch_rewards.append(episode_reward)
                    batch_similarities.append(final_similarity)
                avg_reward = np.mean(batch_rewards)
                avg_sim = np.mean(batch_similarities)
                print(f"    批次 {batch_idx+1}/{num_batches}: 平均奖励={avg_reward:.2f}, 平均相似度={avg_sim:.4f}, epsilon={agent.epsilon:.3f}")
                agent.replay_train(path_idx)
    training_time = time.time() - start_time
    print(f"\n训练完成，总耗时: {training_time:.2f}秒，总步数: {total_steps}")
    dqn_top_k_results = agent.get_all_top_k(EXPERIMENT_CONFIG['TOP_K_SAMPLES'])
    return agent, dqn_top_k_results, training_time, total_steps, agent.replay_train_count

# === 主程序 ===
def main():
    print("=" * 80)
    print("DQN - 20次运行实验")
    print(f"状态空间范围: X[1000,2000], Y[100,200], Z[10,30]")
    print("=" * 80)
    all_dqn_results = []
    all_performance_data = []
    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n开始第 {run_idx+1}/{EXPERIMENT_CONFIG['NUM_RUNS']} 次运行")
        dqn_agent, dqn_results, training_time, total_steps, update_count = train_dqn_workflow()
        performance_data = calculate_run_performance(run_idx, dqn_results, training_time, total_steps, update_count, dqn_agent)
        all_dqn_results.append(dqn_results)
        all_performance_data.append(performance_data)
        print(f"运行 {run_idx+1} 完成: 总奖励={performance_data['Total Reward']}, 收敛度={performance_data['Convergence']}")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"DQN_20_run_{timestamp}.xlsx"
    export_to_excel(all_dqn_results, all_performance_data, targetPaths, output_path)
    print("\n20次运行全部完成！")

if __name__ == "__main__":
    main()