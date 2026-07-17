import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import math

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === 全局取值范围（修改为 X:1000~2000, Y:100~200, Z:10~30） ===
MIN_X = 1000
MAX_X = 2000
MIN_Y = 100
MAX_Y = 200
MIN_Z = 10
MAX_Z = 30

# === 归一化/反归一化 ===
def normalize_state(state):
    """将状态归一化到 [0, 1] 区间"""
    weather_norm = (state[0] - MIN_X) / (MAX_X - MIN_X)
    time_norm = (state[1] - MIN_Y) / (MAX_Y - MIN_Y)
    z_norm = (state[2] - MIN_Z) / (MAX_Z - MIN_Z)
    return (weather_norm, time_norm, z_norm)

def denormalize_state(state_norm):
    """将归一化状态还原"""
    weather = int(round(state_norm[0] * (MAX_X - MIN_X) + MIN_X))
    time_period = int(round(state_norm[1] * (MAX_Y - MIN_Y) + MIN_Y))
    z = int(round(state_norm[2] * (MAX_Z - MIN_Z) + MIN_Z))

    # 边界保护
    weather = np.clip(weather, MIN_X, MAX_X)
    time_period = np.clip(time_period, MIN_Y, MAX_Y)
    z = np.clip(z, MIN_Z, MAX_Z)

    return (weather, time_period, z)

def normalize_value(value, min_val, max_val):
    return (value - min_val) / (max_val - min_val)

def denormalize_value(value_norm, min_val, max_val):
    return int(round(value_norm * (max_val - min_val) + min_val))

# === 安全除法 ===
def safe_divide(numerator, denominator, default=0.0):
    if denominator == 0:
        return default
    return numerator / denominator

# === reward function ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10

    if target_path.issubset(triggered):
        reward += 1

    if prev_triggered is not None:
        prev_sim = jaccard_similarity(prev_triggered, target_path)
        improvement = sim - prev_sim
        reward += improvement * 5

    return reward

# ========== 规则触发函数（section8_hybrid_speed_torque_current） ==========
def section8_hybrid_speed_torque_current(x, y, z):
    """第8类: 速度扭矩电流混合控制 (速度x, 扭矩y, 电流z) - 144个有效变异分支"""
    triggered = set()
    # 删除的100%覆盖率分支: [1, 2, 3, 4, 5, 6] - 不包含在函数中
    # 删除的0%覆盖率分支: [39, 40, 59, 66, 74] - 不包含在函数中
    # 删除缺失分支: [137] (原编号148)

    # -------------------------- 1-32：速度扭矩电流基础检查（删除原1-6，保留7-38） --------------------------

    # 编号1 (原编号7)
    if [(x > 1400 and x < 1600 and y > 160 and y < 190)] != [(x > 1400 and x * 8 < 1600 and y > 160 and y < 190)]:
        triggered.add(1)

    # 编号2 (原编号8)
    if [(x > 1400 and x < 1600 and y > 160 and y < 190)] != [(x > 1400 and x < 1600 and y * 10 > 160 and y < 190)]:
        triggered.add(2)

    # 编号3 (原编号9)
    if [(x > 1350 and x < 1650 and z > 11 and z < 14)] != [(x > 1350 and x * 9 < 1650 and z > 11 and z < 14)]:
        triggered.add(3)

    # 编号4 (原编号10)
    if [(x > 1350 and x < 1650 and z > 11 and z < 14)] != [(x > 1350 and x < 1650 and z * 20 > 11 and z < 14)]:
        triggered.add(4)

    # 编号5 (原编号11)
    if [(y > 155 and y < 195 and z > 11 and z < 14)] != [(y > 155 and y * 7 < 195 and z > 11 and z < 14)]:
        triggered.add(5)

    # 编号6 (原编号12)
    if [(y > 155 and y < 195 and z > 11 and z < 14)] != [(y > 155 and y < 195 and z * 22 > 11 and z < 14)]:
        triggered.add(6)

    # 编号7 (原编号13)
    if [(x > 1480 and x < 1520)] != [(x > 1480 and x * 8 < 1520)]:
        triggered.add(7)

    # 编号8 (原编号14)
    if [(x > 1480 and x < 1520)] != [(x * 12 > 1480 and x < 1520)]:
        triggered.add(8)

    # 编号9 (原编号15)
    if [(y > 173 and y < 177)] != [(y * 12 > 173 and y < 177)]:
        triggered.add(9)

    # 编号10 (原编号16)
    if [(y > 173 and y < 177)] != [(y > 173 and y * 22 < 177)]:
        triggered.add(10)

    # 编号11 (原编号17)
    if [(z > 12.2 and z < 12.8)] != [(z * 78 > 12.2 and z < 12.8)]:
        triggered.add(11)

    # 编号12 (原编号18)
    if [(z > 12.2 and z < 12.8)] != [(66 > 12.2 and z < 12.8)]:
        triggered.add(12)

    # 编号13 (原编号19)
    if [(abs(x - 1500) < 40)] != [(abs(x - 1500) < 140)]:
        triggered.add(13)

    # 编号14 (原编号20)
    if [(abs(x - 1500) < 40)] != [(abs(x - 1500) < 990)]:
        triggered.add(14)

    # 编号15 (原编号21)
    if [(abs(y - 175) < 6)] != [(abs(y - 175) < 16)]:
        triggered.add(15)

    # 编号16 (原编号22)
    if [(abs(y - 175) < 6)] != [(abs(y - 175) < 226)]:
        triggered.add(16)

    # 编号17 (原编号23)
    if [(abs(z - 12.5) < 0.4)] != [(abs(z - 12.5) < 4)]:
        triggered.add(17)

    # 编号18 (原编号24)
    if [(abs(z - 12.5) < 0.4)] != [(abs(z - 12.5) < 1)]:
        triggered.add(18)

    # 编号19 (原编号25)
    if [(x > 1350 and y > 155 and z > 11)] != [(x > 1350 and y * 78 > 155 and z > 11)]:
        triggered.add(19)

    # 编号20 (原编号26)
    if [(x > 1350 and y > 155 and z > 11)] != [(x * 67 > 1350 and y > 155 and z > 11)]:
        triggered.add(20)

    # 编号21 (原编号27)
    if [(x < 1650 and y < 195 and z < 14)] != [(x < 1650 and y * 88 < 195 and z < 14)]:
        triggered.add(21)

    # 编号22 (原编号28)
    if [(x < 1650 and y < 195 and z < 14)] != [(x < 1650 and y < 195 and z * 8 < 14)]:
        triggered.add(22)

    # 编号23 (原编号29)
    if [((x / 10 + y + z * 10) > 350 and (x / 10 + y + z * 10) < 410)] != [
        ((x / 10 + y + z * 10) > 3150 and (x / 10 + y + z * 10) < 410)]:
        triggered.add(23)

    # 编号24 (原编号30)
    if [((x / 10 + y + z * 10) > 350 and (x / 10 + y + z * 10) < 410)] != [
        ((x / 10 + y + z * 10) > 350 and (x / 10 + y + z * 10) < 1410)]:
        triggered.add(24)

    # 编号25 (原编号31)
    if [(x * y / 1000 > 230 and x * y / 1000 < 310)] != [(x * y / 1000 > 1230 and x * y / 1000 < 310)]:
        triggered.add(25)

    # 编号26 (原编号32)
    if [(x * y / 1000 > 230 and x * y / 1000 < 310)] != [(x * y / 1000 > 230 and x * y / 1000 < 3110)]:
        triggered.add(26)

    # 编号27 (原编号33)
    if [(x * z / 100 > 160 and x * z / 100 < 220)] != [(x * z / 100 > 1610 and x * z / 100 < 220)]:
        triggered.add(27)

    # 编号28 (原编号34)
    if [(x * z / 100 > 160 and x * z / 100 < 220)] != [(x * z / 100 > 160 and x * z / 100 < 2120)]:
        triggered.add(28)

    # 编号29 (原编号35)
    if [(y * z > 2000 and y * z < 2600)] != [(y * z > 200 and y * z < 2600)]:
        triggered.add(29)

    # 编号30 (原编号36)
    if [(y * z > 2000 and y * z < 2600)] != [(y * z > 2000 and y * z < 260)]:
        triggered.add(30)

    # 编号31 (原编号37)
    if [((x / 10 + y + z * 10) / 3 > 116 and (x / 10 + y + z * 10) / 3 < 136)] != [
        ((x / 10 + y + z * 10) / 5 > 116 and (x / 10 + y + z * 10) / 3 < 136)]:
        triggered.add(31)

    # 编号32 (原编号38)
    if [((x / 10 + y + z * 10) / 3 > 116 and (x / 10 + y + z * 10) / 3 < 136)] != [
        ((x / 10 + y + z * 10) / 3 > 116 and (x / 10 + y + z * 10) / 11 < 136)]:
        triggered.add(32)

    # 跳过原编号39, 40 (0%覆盖率分支)

    # -------------------------- 33-58：驱动系统分析（原41-80，删除原39、40） --------------------------

    # 编号33 (原编号41)
    if [(x / (y * 10) > 0.8 and x / (y * 10) < 1.0)] != [(x / (y * 10) > 8 and x / (y * 10) < 1.0)]:
        triggered.add(33)

    # 编号34 (原编号42)
    if [(x / (y * 10) > 0.8 and x / (y * 10) < 1.0)] != [(x / (y * 10) > 0.8 and x / (y * 10) < 10)]:
        triggered.add(34)

    # 编号35 (原编号43)
    if [(x / z > 110 and x / z < 140)] != [(x / z > 1220 and x / z < 140)]:
        triggered.add(35)

    # 编号36 (原编号44)
    if [(x / z > 110 and x / z < 140)] != [(x / z > 110 and x / z < 14)]:
        triggered.add(36)

    # 编号37 (原编号45)
    if [(y / z > 12 and y / z < 16)] != [(y / z > 112 and y / z < 16)]:
        triggered.add(37)

    # 编号38 (原编号46)
    if [(y / z > 12 and y / z < 16)] != [(y / z > 12 and y / z < 1226)]:
        triggered.add(38)

    # 编号39 (原编号47)
    if [((x - 1200) / 10 + (y - 150) > 35 and (x - 1200) / 10 + (y - 150) < 55)] != [
        ((x - 1200) / 10 + (y - 150) > 315 and (x - 1200) / 10 + (y - 150) < 55)]:
        triggered.add(39)

    # 编号40 (原编号48)
    if [((x - 1200) / 10 + (y - 150) > 35 and (x - 1200) / 10 + (y - 150) < 55)] != [
        ((x - 1200) / 10 + (y - 220) > 35 and (x - 1200) / 10 + (y - 150) < 55)]:
        triggered.add(40)

    # 编号41 (原编号49)
    if [((x - 1200) / 10 + (z - 10) * 5 > 40 and (x - 1200) / 10 + (z - 10) * 5 < 60)] != [
        ((x - 1200) / 10 + (z - 10) * 15 > 40 and (x - 1200) / 10 + (z - 10) * 5 < 60)]:
        triggered.add(41)

    # 编号42 (原编号50)
    if [((x - 1200) / 10 + (z - 10) * 5 > 40 and (x - 1200) / 10 + (z - 10) * 5 < 60)] != [
        ((x - 1200) / 10 + (z - 10) * 115 > 40 and (x - 1200) / 10 + (z - 10) * 5 < 60)]:
        triggered.add(42)

    # 编号43 (原编号51)
    if [((y - 150) + (z - 10) * 2 > 30 and (y - 150) + (z - 10) * 2 < 50)] != [
        ((y - 150) + (z - 10) * 2 > 302 and (y - 150) + (z - 10) * 2 < 50)]:
        triggered.add(43)

    # 编号44 (原编号52)
    if [((y - 150) + (z - 10) * 2 > 30 and (y - 150) + (z - 10) * 2 < 50)] != [
        ((y - 150) + (z - 10) * 211 > 30 and (y - 150) + (z - 10) * 2 < 50)]:
        triggered.add(44)

    # 编号45 (原编号53)
    if [(abs((x - 1500) / 10 - (y - 175)) < 12)] != [(abs((x - 1500) / 10 - (y - 175)) < 112)]:
        triggered.add(45)

    # 编号46 (原编号54)
    if [(abs((x - 1500) / 10 - (y - 175)) < 12)] != [(abs((x - 1500) / 101 - (y - 175)) < 12)]:
        triggered.add(46)

    # 编号47 (原编号55)
    if [(abs((x - 1500) / 100 - (z - 12.5)) < 3)] != [(abs((x - 1500) / 100 - (z - 12.5)) < 43)]:
        triggered.add(47)

    # 编号48 (原编号56)
    if [(abs((x - 1500) / 100 - (z - 12.5)) < 3)] != [(abs((x - 1500) / 100 - (z - 12.5)) < 13)]:
        triggered.add(48)

    # 编号49 (原编号57)
    if [(abs((y - 175) - (z - 12.5) * 12) < 8)] != [(abs((y - 175) - (z - 12.5) * 12) < 118)]:
        triggered.add(49)

    # 编号50 (原编号58)
    if [(abs((y - 175) - (z - 12.5) * 12) < 8)] != [(abs((y - 175) - (z - 12.5) * 232) < 8)]:
        triggered.add(50)

    # 跳过原编号59 (0%覆盖率分支)

    # 编号51 (原编号60)
    if [(x / (y * 10 + 500) > 0.7 and x / (y * 10 + 500) < 0.9)] != [
        (x / (y * 10 + 500) > 337 and x / (y * 10 + 500) < 0.9)]:
        triggered.add(51)

    # 编号52 (原编号61)
    if [(y / (z + 5) > 9 and y / (z + 5) < 13)] != [(y / (z + 5) > 119 and y / (z + 5) < 13)]:
        triggered.add(52)

    # 编号53 (原编号62)
    if [(y / (z + 5) > 9 and y / (z + 5) < 13)] != [(y / (z + 5) > 9 and y / (z + 5) < 123)]:
        triggered.add(53)

    # 编号54 (原编号63)
    if [(z / (x / 150) > 1.1 and z / (x / 150) < 1.5)] != [(z / (x / 150) > 11 and z / (x / 150) < 1.5)]:
        triggered.add(54)

    # 编号55 (原编号64)
    if [(z / (x / 150) > 1.1 and z / (x / 150) < 1.5)] != [(z / (x / 150) > 1.1 and z / (x / 150) < 15)]:
        triggered.add(55)

    # 编号56 (原编号65)
    if [((x * y * z / 10000) > 28 and (x * y * z / 10000) < 38)] != [
        ((x * y * z / 10000) > 28 and (x * y * z / 10000) < 238)]:
        triggered.add(56)

    # 跳过原编号66 (0%覆盖率分支)

    # 编号57 (原编号67)
    if [(x / 20 + y * 0.6 + z * 8 > 260 and x / 20 + y * 0.6 + z * 8 < 300)] != [
        (x / 20 + y * 0.6 + z * 10 > 260 and x / 20 + y * 0.6 + z * 8 < 300)]:
        triggered.add(57)

    # 编号58 (原编号68)
    if [(x / 20 + y * 0.6 + z * 8 > 260 and x / 20 + y * 0.6 + z * 8 < 300)] != [
        (x / 20 + y * 0.6 + z * 8 > 260 and x / 20 + y * 6 + z * 8 < 300)]:
        triggered.add(58)

    # 编号59 (原编号69)
    if [((x / 1500) ** 0.4 * (y / 175) ** 0.4 * (z / 12.5) ** 0.2 > 0.93)] != [
        ((x / 1500) ** 0.4 * (y / 175) ** 0.4 * (z / 12.5) ** 0.2 > 93)]:
        triggered.add(59)

    # 编号60 (原编号70)
    if [((x / 1500) ** 0.4 * (y / 175) ** 0.4 * (z / 12.5) ** 0.2 > 0.93)] != [
        ((x / 1500) ** 0.4 * (y / 175) ** 4 * (z / 12.5) ** 0.2 > 0.93)]:
        triggered.add(60)

    # 编号61 (原编号71)
    if [((x - 1500) / 10 * (y - 175) > -200 and (x - 1500) / 10 * (y - 175) < 200)] != [
        ((x - 1500) / 110 * (y - 175) > -200 and (x - 1500) / 10 * (y - 175) < 200)]:
        triggered.add(61)

    # 编号62 (原编号72)
    if [((x - 1500) / 10 * (y - 175) > -200 and (x - 1500) / 10 * (y - 175) < 200)] != [
        ((x - 1500) / 10 * (y - 175) > -200 and (x - 1500) / 110 * (y - 175) < 200)]:
        triggered.add(62)

    # 编号63 (原编号73)
    if [((x - 1500) / 100 * (z - 12.5) > -20 and (x - 1500) / 100 * (z - 12.5) < 20)] != [
        ((x - 1500) / 100 * (z - 12.5) > -20 and (x - 1500) / 10 * (z - 12.5) < 20)]:
        triggered.add(63)

    # 跳过原编号74 (0%覆盖率分支)

    # 编号64 (原编号75)
    if [((y - 175) * (z - 12.5) > -20 and (y - 175) * (z - 12.5) < 20)] != [
        ((y - 175) * (z - 12.5) > -20 and (y - 175) * (z - 12.5) < 210)]:
        triggered.add(64)

    # 编号65 (原编号76)
    if [((y - 175) * (z - 12.5) > -20 and (y - 175) * (z - 12.5) < 20)] != [((y - 175) * (z - 125) < 20)]:
        triggered.add(65)

    # 编号66 (原编号77)
    if [(abs(x / 1500 + y / 175 + z / 12.5 - 3) < 0.2)] != [(abs(x / 1500 + y / 175 + z / 12.5 - 3) < 112)]:
        triggered.add(66)

    # 编号67 (原编号78)
    if [(abs(x / 1500 + y / 175 + z / 12.5 - 3) < 0.2)] != [(abs(x / 1500 + y / 175 + z / 12.5 - 3) < 2222)]:
        triggered.add(67)

    # 编号68 (原编号79)
    if [(abs((x / 1500) * (y / 175) * (z / 12.5) - 1) < 0.15)] != [
        (abs((x / 1500) * (y / 175) * (z / 12.5) - 1) < 115)]:
        triggered.add(68)

    # 编号69 (原编号80)
    if [(abs((x / 1500) * (y / 175) * (z / 12.5) - 1) < 0.15)] != [
        (abs((x / 1500) * (y / 175) * (z / 12.5) - 1) < 2225)]:
        triggered.add(69)

    # -------------------------- 70-107：动力传动协调（原81-120） --------------------------

    # 编号70 (原编号81)
    if [(x > 1480 and x < 1520 and y > 172 and y < 178 and z > 12.2 and z < 12.8)] != [
        (x > 1480 and x < 1520 and y > 172 and y < 178 and z > 12.2 and 8 < 12.8)]:
        triggered.add(70)

    # 编号71 (原编号82)
    if [(x > 1480 and x < 1520 and y > 172 and y < 178 and z > 12.2 and z < 12.8)] != [
        (x > 1480 and x < 1520 and y > 172 and y < 178 and z > 12.2 and 10 < 12.8)]:
        triggered.add(71)

    # 编号72 (原编号83)
    if [(abs(x / (y * 10) - 0.86) < 0.06 and abs(z - 12.5) < 0.3)] != [
        (abs(x / (y * 10) - 0.86) < 0.06 and abs(z - 12.5) < 3)]:
        triggered.add(72)

    # 编号73 (原编号84)
    if [(abs(x / (y * 10) - 0.86) < 0.06 and abs(z - 12.5) < 0.3)] != [
        (abs(x / (y * 10) - 0.86) < 0.06 and abs(z - 12.5) < 13)]:
        triggered.add(73)

    # 编号74 (原编号85)
    if [(x * y / 1000 > 258 and x * y / 1000 < 268 and z > 12.3 and z < 12.7)] != [
        (x * y / 1000 > 258 and x * y / 1000 < 268 and 8 * z > 12.3 and z < 12.7)]:
        triggered.add(74)

    # 编号75 (原编号86)
    if [(x * y / 1000 > 258 and x * y / 1000 < 268 and z > 12.3 and z < 12.7)] != [
        (x * y / 1000 > 258 and x * y / 1000 < 268 and z * 99 > 12.3 and z < 12.7)]:
        triggered.add(75)

    # 编号76 (原编号87)
    if [((x / 10 + y) / 2 > 137 and (x / 10 + y) / 2 < 143 and z > 12.2 and z < 12.8)] != [
        ((x / 10 + y) / 2 > 137 and (x / 10 + y) / 2 < 143 and z > 12.2 and 8 < 12.8)]:
        triggered.add(76)

    # 编号77 (原编号88)
    if [((x / 10 + y) / 2 > 137 and (x / 10 + y) / 2 < 143 and z > 12.2 and z < 12.8)] != [
        ((x / 10 + y) / 2 > 137 and (x / 10 + y) / 2 < 143 and z > 12.2 and 10 < 12.8)]:
        triggered.add(77)

    # 编号78 (原编号89)
    if [(abs(x / 10 - y) < 25 and z > 12.2 and z < 12.8)] != [(abs(x / 10 - y) < 25 and z * 12 > 12.2 and z < 12.8)]:
        triggered.add(78)

    # 编号79 (原编号90)
    if [(abs(x / 10 - y) < 25 and z > 12.2 and z < 12.8)] != [(abs(x / 10 - y) < 25 and z * 89 > 12.2 and z < 12.8)]:
        triggered.add(79)

    # 编号80 (原编号91)
    if [(math.sqrt((x / 10 - 150) ** 2 + (y - 175) ** 2) < 6 and abs(z - 12.5) < 0.3)] != [
        (math.sqrt((x / 10 - 150) ** 2 + (y - 175) ** 2) < 6 and abs(z - 12.5) < 3)]:
        triggered.add(80)

    # 编号81 (原编号92)
    if [(math.sqrt((x / 10 - 150) ** 2 + (y - 175) ** 2) < 6 and abs(z - 12.5) < 0.3)] != [
        (math.sqrt((x / 10 - 150) ** 2 + (y - 175) ** 2) < 6 and abs(z - 12.5) < 13)]:
        triggered.add(81)

    # 编号82 (原编号93)
    if [(x / 1500 > 0.98 and x / 1500 < 1.02 and y / 175 > 0.97 and y / 175 < 1.03)] != [
        (x / 1500 > 0.98 and x / 1500 < 1.02 and y / 175 > 97 and y / 175 < 1.03)]:
        triggered.add(82)

    # 编号83 (原编号94)
    if [(x / 1500 > 0.98 and x / 1500 < 1.02 and y / 175 > 0.97 and y / 175 < 1.03)] != [
        (x / 1500 > 0.98 and x / 1500 < 1.02 and y * 8 / 175 > 0.97 and y / 175 < 1.03)]:
        triggered.add(83)

    # 编号84 (原编号95)
    if [(z / 12.5 > 0.976 and z / 12.5 < 1.024)] != [(z / 12.5 > 0.976 and z / 125 < 1.024)]:
        triggered.add(84)

    # 编号85 (原编号96)
    if [(z / 12.5 > 0.976 and z / 12.5 < 1.024)] != [(z / 12.5 > 0.976 and z / 200 < 1.024)]:
        triggered.add(85)

    # 编号86 (原编号97)
    if [((x / 1500 + y / 175 + z / 12.5) / 3 > 0.98 and (x / 1500 + y / 175 + z / 12.5) / 3 < 1.02)] != [
        ((x / 1500 + y / 175 + z / 12.5) / 3 > 98 and (x / 1500 + y / 175 + z / 12.5) / 3 < 1.02)]:
        triggered.add(86)

    # 编号87 (原编号98)
    if [((x / 1500 + y / 175 + z / 12.5) / 3 > 0.98 and (x / 1500 + y / 175 + z / 12.5) / 3 < 1.02)] != [
        ((x / 1500 + y / 175 + z / 12.5) / 23 > 0.98 and (x / 1500 + y / 175 + z / 12.5) / 3 < 1.02)]:
        triggered.add(87)

    # 编号88 (原编号99)
    if [(max(abs(x / 1500 - 1), abs(y / 175 - 1), abs(z / 12.5 - 1)) < 0.025)] != [
        (max(abs(x / 1500 - 1), abs(y / 175 - 1), abs(z / 12.5 - 1)) < 0.55)]:
        triggered.add(88)

    # 编号89 (原编号100)
    if [(max(abs(x / 1500 - 1), abs(y / 175 - 1), abs(z / 12.5 - 1)) < 0.025)] != [
        (max(abs(x / 1500 - 1), abs(y / 175 - 1), abs(z / 12.5 - 1)) < 0.25)]:
        triggered.add(89)

    # 编号90 (原编号101)
    if [(min(x / 1500, y / 175, z / 12.5) > 0.98)] != [(min(x / 1500, y / 175, z / 12.5) > 498)]:
        triggered.add(90)

    # 编号91 (原编号102)
    if [(min(x / 1500, y / 175, z / 12.5) > 0.98)] != [(min(x / 1500, y / 1375, z / 12.5) > 0.98)]:
        triggered.add(91)

    # 编号92 (原编号103)
    if [(max(x / 1500, y / 175, z / 12.5) < 1.02)] != [(max(x / 1500, y / 175, z / 12.5) < 302)]:
        triggered.add(92)

    # 编号93 (原编号104)
    if [(max(x / 1500, y / 175, z / 12.5) < 1.02)] != [(max(x / 1500, y / 175, z / 12.5) < 102)]:
        triggered.add(93)

    # 编号94 (原编号105)
    if [(abs(max(x / 10, y, z * 10) - min(x / 10, y, z * 10)) < 30)] != [
        (abs(max(x / 10, y, z * 120) - min(x / 10, y, z * 10)) < 30)]:
        triggered.add(94)

    # 编号95 (原编号106)
    if [(abs(max(x / 10, y, z * 10) - min(x / 10, y, z * 10)) < 30)] != [
        (abs(max(x / 10, y, z * 10) - min(x / 10, y, z * 10)) < 230)]:
        triggered.add(95)

    # 编号96 (原编号107)
    if [((x / 10 + y + z * 10) > 378 and (x / 10 + y + z * 10) < 382)] != [
        ((x / 10 + y + z * 10) > 378 and (x / 10 + y + z * 10) < 3112)]:
        triggered.add(96)

    # 编号97 (原编号108)
    if [((x / 10 + y + z * 10) > 378 and (x / 10 + y + z * 10) < 382)] != [
        ((x / 10 + y + z * 10) > 3781 and (x / 10 + y + z * 10) < 382)]:
        triggered.add(97)

    # 编号98 (原编号109)
    if [((x * y * z / 10000) > 32.5 and (x * y * z / 10000) < 33.5)] != [
        ((x * y * z / 10000) > 32.5 and (x * y * 2 / 10000) < 33.5)]:
        triggered.add(98)

    # 编号99 (原编号110)
    if [((x * y * z / 10000) > 32.5 and (x * y * z / 10000) < 33.5)] != [
        ((x * y * z / 10000) > 32.5 and (x * y * 2.4 / 10000) < 33.5)]:
        triggered.add(99)

    # 编号100 (原编号111)
    if [(abs((x / 10 + y + z * 10) / 3 - 126.7) < 1.5)] != [(abs((x / 10 + y + z * 10) / 3 - 126.7) < 15)]:
        triggered.add(100)

    # 编号101 (原编号112)
    if [(abs((x / 10 + y + z * 10) / 3 - 126.7) < 1.5)] != [(abs((x / 10 + y + z * 10) / 3 - 126.7) < 115)]:
        triggered.add(101)

    # 编号102 (原编号113)
    if [(x > 1495 and x < 1505 and y > 174 and y < 176)] != [(x > 1495 and x < 1505 and y * 10 > 174 and y < 176)]:
        triggered.add(102)

    # 编号103 (原编号114)
    if [(x > 1495 and x < 1505 and y > 174 and y < 176)] != [(x > 1495 and x < 1505 and y * 88 > 174 and y < 176)]:
        triggered.add(103)

    # 编号104 (原编号115)
    if [(z > 12.45 and z < 12.55)] != [(z * 10 > 12.45 and z < 12.55)]:
        triggered.add(104)

    # 编号105 (原编号116)
    if [(z > 12.45 and z < 12.55)] != [(z * 8 > 12.45 and z < 12.55)]:
        triggered.add(105)

    # -------------------------- 106-144：动力控制优化（原117-156，删除原148） --------------------------

    # 编号106 (原编号117)
    if [(x < 1350 or y < 160 or z < 11.5)] != [(x < 1350 or y * 8 < 160 or z < 11.5)]:
        triggered.add(106)

    # 编号107 (原编号118)
    if [(x < 1350 or y < 160 or z < 11.5)] != [(x < 1350 or y < 160 or z * 89 < 11.5)]:
        triggered.add(107)

    # 编号108 (原编号119)
    if [(x > 1650 or y > 190 or z > 13.5)] != [(x > 1650 or y > 190 or z > 135)]:
        triggered.add(108)

    # 编号109 (原编号120)
    if [(x > 1650 or y > 190 or z > 13.5)] != [(x > 1650 or y * 87 > 190 or z > 13.5)]:
        triggered.add(109)

    # 编号110 (原编号121)
    if [(abs(x / (y * 10) - 0.86) > 0.1)] != [(abs(x / (y * 10) - 0.86) > 111)]:
        triggered.add(110)

    # 编号111 (原编号122)
    if [(abs(x / (y * 10) - 0.86) > 0.1)] != [(abs(x / (y * 10) - 0.86) > 1671)]:
        triggered.add(111)

    # 编号112 (原编号123)
    if [(abs(x / z - 120) > 15)] != [(abs(x / z - 120) > 1522)]:
        triggered.add(112)

    # 编号113 (原编号124)
    if [(abs(x / z - 120) > 15)] != [(abs(x / z - 120) > 115)]:
        triggered.add(113)

    # 编号114 (原编号125)
    if [(abs(y / z - 14) > 2)] != [(abs(y / z - 14) > 112)]:
        triggered.add(114)

    # 编号115 (原编号126)
    if [(abs(y / z - 14) > 2)] != [(abs(y / z - 14) > 211)]:
        triggered.add(115)

    # 编号116 (原编号127)
    if [((x / 10 + y + z * 10) < 360 or (x / 10 + y + z * 10) > 400)] != [
        ((x / 10 + y + z * 10) < 36 or (x / 10 + y + z * 10) > 400)]:
        triggered.add(116)

    # 编号117 (原编号128)
    if [((x / 10 + y + z * 10) < 360 or (x / 10 + y + z * 10) > 400)] != [
        ((x / 10 + y + z * 10) < 989 or (x / 10 + y + z * 10) > 400)]:
        triggered.add(117)

    # 编号118 (原编号129)
    if [((x * y * z / 10000) < 30 or (x * y * z / 10000) > 36)] != [
        ((x * y * z / 10000) < 30 or (x * y * z / 10000) > 136)]:
        triggered.add(118)

    # 编号119 (原编号130)
    if [((x * y * z / 10000) < 30 or (x * y * z / 10000) > 36)] != [
        ((x * y * z / 10000) < 30 or (x * y * z / 10000) > 316)]:
        triggered.add(119)

    # 编号120 (原编号131)
    if [(x < 1300 and y < 160)] != [(x < 1300 and y * 878 < 160)]:
        triggered.add(120)

    # 编号121 (原编号132)
    if [(x < 1300 and y < 160)] != [(x * 71 < 1300 and y < 160)]:
        triggered.add(121)

    # 编号122 (原编号133)
    if [(x > 1700 and y > 190)] != [(x > 1700 and y * 78 > 190)]:
        triggered.add(122)

    # 编号123 (原编号134)
    if [(x > 1700 and y > 190)] != [(x * 78 > 1700 and y > 190)]:
        triggered.add(123)

    # 编号124 (原编号135)
    if [(x < 1300 and z < 11)] != [(x < 1300 and z * 91 < 11)]:
        triggered.add(124)

    # 编号125 (原编号136)
    if [(x < 1300 and z < 11)] != [(x * 12 < 1300 and z < 11)]:
        triggered.add(125)

    # 编号126 (原编号137)
    if [(x > 1700 and z > 14)] != [(x > 1700 and z * 21 > 14)]:
        triggered.add(126)

    # 编号127 (原编号138)
    if [(x > 1700 and z > 14)] != [(x * 123 > 1700 and z > 14)]:
        triggered.add(127)

    # 编号128 (原编号139)
    if [(y < 160 and z < 11)] != [(y < 160 and z * 78 < 11)]:
        triggered.add(128)

    # 编号129 (原编号140)
    if [(y < 160 and z < 11)] != [(y * 8 < 160 and z < 11)]:
        triggered.add(129)

    # 编号130 (原编号141)
    if [(y > 190 and z > 14)] != [(y * 8 > 190 and z > 14)]:
        triggered.add(130)

    # 编号131 (原编号142)
    if [(y > 190 and z > 14)] != [(y > 190 and z * 9 > 14)]:
        triggered.add(131)

    # 编号132 (原编号143)
    if [(x < 1250 or x > 1750)] != [(x * 67 < 1250 or x > 1750)]:
        triggered.add(132)

    # 编号133 (原编号144)
    if [(x < 1250 or x > 1750)] != [(x < 1250 or x * 53 > 1750)]:
        triggered.add(133)

    # 编号134 (原编号145)
    if [(y < 145 or y > 205)] != [(y * 67 < 145 or y > 205)]:
        triggered.add(134)

    # 编号135 (原编号146)
    if [(y < 145 or y > 205)] != [(y < 145 or y * 67 > 205)]:
        triggered.add(135)

    # 编号136 (原编号147)
    if [(z < 9.5 or z > 15.5)] != [(z < 9.5 or z * 67 > 15.5)]:
        triggered.add(136)

    # 跳过原编号148 (缺失分支)

    # 编号137 (原编号149)
    if [(x < 1200 and y < 155 and z < 10.5)] != [(x * 67 < 1200 and y < 155 and z < 10.5)]:
        triggered.add(137)

    # 编号138 (原编号150)
    if [(x < 1200 and y < 155 and z < 10.5)] != [(x < 1200 and y < 55 and z < 10.5)]:
        triggered.add(138)

    # 编号139 (原编号151)
    if [(x > 1800 and y > 195 and z > 14.5)] != [(x > 1800 and y * 78 > 195 and z > 14.5)]:
        triggered.add(139)

    # 编号140 (原编号152)
    if [(x > 1800 and y > 195 and z > 14.5)] != [(x > 1800 and y > 966 and z > 14.5)]:
        triggered.add(140)

    # 编号141 (原编号153)
    if [((x / 10 - 150) ** 2 + (y - 175) ** 2 + (z - 12.5) ** 2 * 100) > 150] != [
        ((x / 10 - 150) ** 6 + (y - 175) ** 2 + (z - 12.5) ** 2 * 100) > 150]:
        triggered.add(141)

    # 编号142 (原编号154)
    if [((x / 10 - 150) ** 2 + (y - 175) ** 2 + (z - 12.5) ** 2 * 100) > 150] != [
        ((x / 10 - 150) ** 2 + (y - 175) ** 8 + (z - 12.5) ** 2 * 100) > 150]:
        triggered.add(142)

    # 编号143 (原编号155)
    if [(abs((x * y * z / 10000) / 33 - 1) > 0.18)] != [(abs((x * y * z / 10000) / 33 - 1) > 18)]:
        triggered.add(143)

    # 编号144 (原编号156)
    if [(abs((x * y * z / 10000) / 33 - 1) > 0.18)] != [(abs((x * y * z / 10000) / 33 - 1) > 10)]:
        triggered.add(144)

    return triggered

# === 目标路径组（section8 的编号 1~144） ===
targetPaths = [
    {1, 4, 6, 9, 11, 12, 13, 14, 15, 16, 17, 21, 22, 24, 25, 27, 29, 32, 33, 38, 44, 45, 46, 47, 48, 49, 51, 52, 58, 59,
     60, 64, 68, 69, 74, 75, 78, 79, 88, 89, 92, 93, 95, 96, 101, 104, 105, 107, 109, 110, 111, 112, 113, 114, 115, 119,
     133, 135, 136, 143, 144},

    {4, 6, 9, 11, 12, 13, 14, 15, 16, 17, 21, 22, 24, 25, 27, 29, 32, 34, 37, 39, 40, 44, 45, 47, 48, 50, 51, 52, 58,
     59, 60, 61, 64, 68, 69, 74, 75, 78, 79, 88, 89, 92, 93, 95, 96, 101, 104, 105, 109, 110, 111, 112, 113, 119, 128,
     129, 133, 135, 136, 143, 144},

    {3, 5, 9, 11, 12, 13, 14, 16, 17, 18, 21, 22, 24, 25, 27, 29, 32, 34, 35, 36, 37, 39, 40, 44, 45, 49, 51, 52, 54,
     58, 59, 60, 61, 65, 74, 75, 78, 79, 86, 87, 88, 89, 92, 93, 95, 96, 101, 104, 105, 106, 109, 110, 111, 112, 113,
     119, 133, 135, 136, 143, 144},

    {2, 3, 8, 9, 11, 12, 13, 14, 16, 17, 18, 19, 21, 22, 23, 27, 29, 31, 34, 35, 36, 41, 42, 44, 45, 49, 51, 54, 56, 57,
     62, 64, 66, 67, 68, 69, 78, 79, 88, 89, 94, 96, 100, 101, 104, 105, 106, 109, 110, 111, 114, 115, 117, 119, 133,
     134, 136, 143, 144},

    {2, 4, 7, 9, 11, 12, 16, 17, 21, 22, 24, 27, 29, 32, 33, 35, 36, 37, 41, 42, 44, 45, 50, 51, 52, 57, 64, 66, 67, 68,
     69, 78, 79, 83, 88, 89, 95, 96, 100, 101, 102, 103, 104, 105, 109, 110, 111, 112, 113, 119, 128, 129, 133, 135,
     136, 143, 144},

    {6, 8, 11, 12, 14, 15, 16, 17, 21, 22, 24, 29, 32, 35, 36, 38, 39, 40, 42, 43, 45, 49, 52, 54, 56, 57, 60, 61, 63,
     65, 66, 67, 68, 69, 88, 89, 92, 93, 95, 96, 100, 101, 104, 105, 109, 110, 111, 114, 115, 119, 124, 125, 132, 135,
     136, 143, 144},

    {1, 4, 6, 8, 9, 11, 12, 13, 14, 15, 16, 17, 21, 22, 24, 25, 29, 32, 33, 35, 36, 37, 39, 40, 42, 44, 49, 52, 54, 57,
     59, 60, 65, 66, 67, 68, 69, 72, 73, 78, 79, 88, 89, 95, 96, 100, 101, 104, 105, 107, 109, 119, 133, 135, 136, 143,
     144},

    {1, 3, 5, 7, 9, 11, 12, 17, 18, 21, 22, 24, 25, 27, 30, 32, 33, 35, 36, 37, 39, 40, 41, 42, 44, 50, 52, 54, 58, 59,
     65, 72, 73, 74, 75, 78, 79, 80, 81, 82, 88, 89, 95, 96, 101, 102, 103, 104, 105, 109, 119, 133, 135, 136, 142, 143,
     144},

    {8, 9, 11, 12, 14, 16, 17, 21, 22, 23, 29, 31, 33, 35, 36, 37, 44, 46, 49, 54, 56, 62, 63, 64, 66, 67, 68, 69, 72,
     73, 78, 79, 88, 89, 95, 98, 100, 101, 104, 105, 109, 117, 119, 120, 121, 124, 125, 128, 129, 132, 134, 136, 137,
     138, 143, 144},

    {9, 11, 12, 14, 16, 17, 24, 25, 27, 29, 32, 34, 37, 39, 40, 45, 47, 48, 50, 51, 52, 57, 59, 60, 61, 64, 68, 69, 74,
     75, 78, 79, 88, 89, 92, 93, 95, 96, 100, 101, 104, 105, 110, 111, 112, 113, 119, 122, 126, 128, 129, 133, 135, 136,
     143, 144},

    {6, 8, 11, 12, 14, 16, 17, 21, 22, 23, 29, 31, 35, 36, 38, 39, 40, 43, 45, 49, 52, 54, 56, 57, 60, 61, 63, 65, 66,
     67, 68, 69, 88, 89, 92, 93, 95, 96, 100, 101, 104, 105, 110, 111, 114, 115, 119, 123, 124, 125, 131, 132, 135, 136,
     143, 144},

    {8, 9, 11, 12, 14, 16, 17, 21, 22, 29, 34, 35, 36, 42, 44, 45, 49, 51, 54, 56, 62, 63, 64, 66, 67, 68, 69, 78, 79,
     88, 94, 98, 99, 100, 101, 104, 105, 109, 110, 111, 114, 115, 116, 119, 120, 121, 124, 125, 128, 129, 133, 134, 136,
     143, 144},

    {5, 8, 11, 12, 14, 16, 17, 20, 21, 22, 24, 30, 32, 38, 39, 40, 42, 43, 45, 49, 52, 54, 58, 60, 61, 63, 65, 66, 67,
     68, 69, 88, 89, 92, 93, 95, 96, 100, 101, 104, 105, 110, 111, 112, 113, 114, 115, 119, 123, 131, 132, 135, 136,
     143, 144},

    {8, 9, 14, 16, 17, 23, 27, 29, 31, 34, 41, 42, 44, 45, 47, 48, 49, 51, 55, 56, 57, 62, 66, 67, 68, 69, 84, 85, 88,
     92, 93, 95, 96, 98, 99, 100, 101, 108, 110, 111, 112, 113, 114, 115, 117, 119, 120, 121, 127, 130, 132, 134, 136,
     143, 144},

    {5, 8, 9, 14, 15, 16, 17, 18, 20, 21, 22, 23, 30, 32, 37, 41, 42, 44, 45, 46, 47, 48, 49, 52, 55, 58, 62, 65, 66,
     67, 68, 69, 76, 77, 84, 85, 88, 89, 92, 93, 95, 96, 100, 101, 109, 110, 111, 112, 113, 119, 132, 135, 136, 143,
     144},

    {1, 3, 5, 9, 11, 12, 17, 18, 21, 22, 24, 25, 27, 30, 32, 33, 35, 36, 37, 39, 40, 44, 50, 52, 54, 58, 59, 60, 65, 72,
     73, 74, 75, 78, 79, 80, 81, 82, 88, 89, 95, 96, 101, 104, 105, 109, 119, 133, 135, 136, 141, 142, 143, 144},

    {1, 4, 6, 8, 10, 11, 12, 17, 21, 22, 24, 25, 29, 32, 33, 35, 36, 38, 39, 40, 42, 44, 49, 52, 58, 59, 68, 69, 72, 73,
     74, 75, 80, 81, 82, 88, 89, 95, 96, 101, 104, 105, 107, 109, 112, 113, 114, 115, 119, 133, 135, 136, 143, 144},

    {8, 9, 11, 12, 14, 16, 17, 21, 22, 23, 29, 31, 33, 35, 36, 37, 42, 44, 46, 50, 52, 54, 56, 62, 63, 64, 66, 67, 68,
     69, 72, 73, 78, 79, 88, 89, 95, 97, 104, 105, 109, 117, 119, 120, 121, 124, 125, 128, 129, 133, 135, 136, 143,
     144},

    {8, 9, 11, 12, 14, 16, 17, 21, 22, 29, 34, 42, 44, 45, 49, 51, 54, 56, 63, 64, 66, 67, 68, 69, 78, 79, 88, 94, 98,
     99, 101, 104, 105, 109, 110, 111, 112, 113, 114, 115, 116, 118, 119, 120, 121, 124, 125, 128, 129, 132, 134, 136,
     143, 144},

    {5, 11, 12, 14, 15, 16, 17, 24, 26, 27, 30, 32, 33, 38, 43, 45, 46, 47, 48, 49, 51, 52, 58, 59, 62, 66, 67, 68, 69,
     78, 79, 88, 89, 92, 93, 95, 96, 101, 104, 105, 107, 110, 111, 112, 113, 114, 115, 122, 126, 135, 136, 143},

    {1, 3, 5, 7, 9, 21, 22, 24, 25, 27, 30, 32, 33, 35, 36, 37, 39, 40, 44, 50, 52, 54, 58, 59, 65, 70, 71, 72, 73, 80,
     81, 82, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 95, 96, 101, 109, 133, 135, 136, 142, 143, 144},

    {2, 8, 9, 13, 14, 16, 17, 19, 24, 28, 30, 32, 34, 44, 45, 47, 48, 49, 51, 55, 58, 59, 60, 62, 76, 77, 84, 85, 86,
     87, 88, 92, 93, 94, 96, 101, 106, 108, 110, 111, 112, 113, 114, 115, 119, 127, 130, 133, 134, 143, 144},

    {14, 15, 16, 17, 24, 26, 28, 32, 33, 35, 36, 37, 43, 45, 46, 49, 51, 52, 54, 59, 62, 63, 64, 65, 66, 67, 68, 69, 84,
     85, 88, 89, 90, 91, 92, 93, 95, 96, 101, 110, 111, 122, 130, 135, 136, 139, 143},

    {11, 12, 14, 16, 17, 24, 26, 27, 29, 32, 33, 38, 43, 46, 47, 48, 49, 51, 53, 58, 59, 62, 65, 72, 73, 78, 79, 88, 89,
     92, 93, 95, 96, 101, 104, 105, 107, 112, 113, 114, 115, 126, 131, 135, 136, 143, 144},

    {14, 16, 17, 24, 26, 28, 32, 33, 35, 36, 37, 45, 49, 51, 52, 54, 59, 62, 63, 64, 65, 66, 67, 68, 69, 84, 85, 88, 90,
     91, 92, 93, 95, 96, 101, 110, 111, 135, 136, 140, 143},
]

# 将路径列表转换为集合列表
target_paths = [set(path) for path in targetPaths]
NUM_PATHS = len(target_paths)

def jaccard_similarity(set1, set2):
    """Compute Jaccard similarity"""
    if not set2:
        return 0.0

    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0

# === Path Similarity ===
def compute_path_similarity_matrix(paths):
    """Path Similarity"""
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix

def group_paths_by_similarity(paths, threshold_percentile=50):
    """SimilarityPath """
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.percentile(avg_sim_scores, threshold_percentile)

    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]

    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)

    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group

# === Sample generation ===
def compute_robustness(state, path, sample_size=9):
    """计算鲁棒性"""
    base = section8_hybrid_speed_torque_current(state[0], state[1], state[2])
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0

    # 邻域偏移量（z范围10~30，步长适当）
    deltas = [
        (-1, -1, -4), (0, -1, 0), (1, -1, 4),
        (-1, 0, -4), (1, 0, 4),
        (-1, 1, -4), (0, 1, 0), (1, 1, 4),
        (0, 0, 0)
    ]

    for dw, dt, dz in deltas[:sample_size]:
        if dw == dt == dz == 0:
            continue

        neighbor_weather = int(np.clip(state[0] + dw, MIN_X, MAX_X))
        neighbor_time = int(np.clip(state[1] + dt, MIN_Y, MAX_Y))
        neighbor_z = int(np.clip(state[2] + dz, MIN_Z, MAX_Z))
        neighbor = (neighbor_weather, neighbor_time, neighbor_z)

        n_trig = section8_hybrid_speed_torque_current(neighbor[0], neighbor[1], neighbor[2])
        if not n_trig:
            continue

        rob += jaccard_similarity(n_trig, base)
        neighbors += 1

    return rob / neighbors if neighbors > 0 else 0.0

def generate_samples_for_all_paths(num_candidates=2000, top_k=200, run_id=1):
    """为所有路径生成样本（权重筛选）"""
    BEST_WEIGHTS = [0.55, 0.25, 0.2]

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_individual.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Individual Path {path_id} (Weighted Screening) - Run {run_id}\n")
            f.write("weather time_period z\tScore\tSimilarity\tRobustness\tLengthDiff\n")
            for s in samples:
                weather, time_period, z = s['state']
                f.write(
                    f"{weather} {time_period} {z}\t{s['score']:.4f}\t{s['similarity']:.4f}\t"
                    f"{s['robustness']:.4f}\t{s['length_diff']:.4f}\n"
                )

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_individual"

    for path_idx in range(len(target_paths)):
        path = target_paths[path_idx]
        candidate_samples = []
        attempts = 0
        max_attempts = num_candidates * 10

        while len(candidate_samples) < num_candidates and attempts < max_attempts:
            attempts += 1

            weather = np.random.randint(MIN_X, MAX_X + 1)
            time_period = np.random.randint(MIN_Y, MAX_Y + 1)
            z = np.random.randint(MIN_Z, MAX_Z + 1)
            state = (weather, time_period, z)

            triggered = section8_hybrid_speed_torque_current(weather, time_period, z)
            if not triggered:
                continue

            sim = jaccard_similarity(triggered, path)
            rob = compute_robustness(state, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))

            candidate_samples.append({
                'state': state,
                'similarity': sim,
                'robustness': rob,
                'length_diff': len_diff,
                'triggered': triggered
            })

        if candidate_samples:
            for sample in candidate_samples:
                score = (BEST_WEIGHTS[0] * sample['similarity'] +
                         BEST_WEIGHTS[1] * sample['robustness'] +
                         BEST_WEIGHTS[2] * sample['length_diff'])
                sample['score'] = score

            candidate_samples.sort(key=lambda x: x['score'], reverse=True)
            selected_samples = candidate_samples[:top_k]
            save_samples(path_id=path_idx + 1, samples=selected_samples, base_dir=base_dir)

# === Shared Experience Replay ===
class SharedExperienceReplay:
    def __init__(self, capacity=20000):
        self.capacity = capacity
        self.buffer = deque(maxlen=self.capacity)
        self.priorities = deque(maxlen=self.capacity)

    def append(self, experience):
        self.buffer.append(experience)
        self.priorities.append(experience[-1])

    def sample(self, batch_size, alpha=0.6):
        if len(self.buffer) < batch_size:
            return [], [], []

        priorities = np.array(self.priorities, dtype=np.float64)
        priorities = np.power(priorities, alpha)
        probabilities = priorities / np.sum(priorities)

        batch_indices = np.random.choice(len(self.buffer), batch_size, p=probabilities, replace=False)
        batch = [self.buffer[idx] for idx in batch_indices]

        return batch, batch_indices, probabilities[batch_indices]

    def update_priorities(self, batch_indices, td_errors):
        for idx, td_error in zip(batch_indices, td_errors):
            if idx < len(self.priorities):
                self.priorities[idx] = max(abs(td_error), 1e-6)

    def __len__(self):
        return len(self.buffer)

    def get_high_reward_samples(self, target_path, num_samples=20):
        if len(self.buffer) == 0:
            return []

        samples_with_scores = []
        seen_states = set()

        for experience in self.buffer:
            state_tensor = experience[0]
            state_norm = state_tensor.cpu().numpy().flatten()
            state_tuple = denormalize_state((state_norm[0], state_norm[1], state_norm[2]))

            if state_tuple in seen_states:
                continue
            seen_states.add(state_tuple)

            triggered = section8_hybrid_speed_torque_current(state_tuple[0], state_tuple[1], state_tuple[2])
            reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)

            samples_with_scores.append((state_tuple, reward, sim, triggered))

        samples_with_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_scores[:num_samples]

def load_path_data(file_path):
    path_data = []

    if not os.path.exists(file_path):
        print(f"文件不存在: {file_path}")
        return path_data

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
            for line in lines[2:]:
                parts = line.strip().split("\t")
                if parts:
                    values = parts[0].split()
                    if len(values) >= 3:
                        state = (int(values[0]), int(values[1]), int(values[2]))
                        path_data.append(state)
    except Exception as e:
        print(f"读取文件失败 {file_path}: {e}")

    return path_data

# === DQN ===
class DQN(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dims=[128, 64]):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, hidden_dims[0])
        self.fc2 = nn.Linear(hidden_dims[0], hidden_dims[1])
        self.fc3 = nn.Linear(hidden_dims[1], action_dim)
        self.dropout = nn.Dropout(0.1)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = self.dropout(x)
        x = torch.relu(self.fc2(x))
        return self.fc3(x)

# === DQN Agent with PER ===
class DQNAgentWithPER:
    def __init__(self, state_dim, action_dim, replay_buffer,
                 gamma=0.99, epsilon=1.0, epsilon_decay=0.995,
                 epsilon_min=0.1, learning_rate=0.001, alpha=0.6, beta=0.4):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer
        self.alpha = alpha
        self.beta = beta

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        """
        动作解码：30个动作，分别对应三个维度的不同步长
        - weather: +/-1, 0(x2)
        - time_period: +/-1, 0(x2)
        - z: +/-4, +/-2, +/-1, 0(x2)  适配 10~30 范围
        """
        delta_values_weather_time = [1, 0, 0, -1]
        delta_values_z = [4, 2, 1, 0, 0, -1, -2, -4]

        dim = action_idx // 10
        delta_idx = action_idx % 10

        if dim == 0:  # weather
            if delta_idx >= 4:
                delta_idx = 3
            return (delta_values_weather_time[delta_idx], 0, 0)
        elif dim == 1:  # time_period
            if delta_idx >= 4:
                delta_idx = 3
            return (0, delta_values_weather_time[delta_idx], 0)
        elif dim == 2:  # z
            if delta_idx >= 8:
                delta_idx = 7
            return (0, 0, delta_values_z[delta_idx])

    def act(self, state_norm, legal_actions=None):
        if legal_actions is None:
            legal_actions = list(range(self.action_dim))

        if not legal_actions:
            return None

        if random.random() < self.epsilon:
            return random.choice(legal_actions)

        state_tensor = torch.tensor(state_norm, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)[0]

        legal_q_values = q_values[legal_actions]
        best_legal_idx = torch.argmax(legal_q_values).item()
        return legal_actions[best_legal_idx]

    def get_legal_actions(self, state):
        legal_actions = []

        for action_idx in range(self.action_dim):
            dw, dt, dz = self.decode_action(action_idx)

            next_weather = state[0] + dw
            next_time = state[1] + dt
            next_z = state[2] + dz

            if (MIN_X <= next_weather <= MAX_X and
                    MIN_Y <= next_time <= MAX_Y and
                    MIN_Z <= next_z <= MAX_Z):
                legal_actions.append(action_idx)

        return legal_actions

    def store_transition(self, state, action, reward, next_state, done):
        state_norm = normalize_state(state)
        next_state_norm = normalize_state(next_state)

        state_tensor = torch.tensor(state_norm, dtype=torch.float32).unsqueeze(0).to(device)
        next_state_tensor = torch.tensor(next_state_norm, dtype=torch.float32).unsqueeze(0).to(device)

        with torch.no_grad():
            q_values = self.model(state_tensor)
            next_q_values = self.target_model(next_state_tensor)
            max_next_q = next_q_values.max(1)[0]
            target_q = reward + (self.gamma * max_next_q * (1 - done))
            td_error = abs(q_values[0][action].item() - target_q.item())

        self.replay_buffer.append((state_tensor, action, reward, next_state_tensor, done, td_error))
        return td_error

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return 0.0

        batch, batch_indices, probabilities = self.replay_buffer.sample(batch_size, alpha=self.alpha)

        if not batch:
            return 0.0

        states, actions, rewards, next_states, dones, _ = zip(*batch)

        weights = (len(self.replay_buffer) * probabilities) ** (-self.beta)
        weights = weights / weights.max()
        weights = torch.tensor(weights, dtype=torch.float32).to(device)

        states = torch.cat(states).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.cat(next_states).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q = self.target_model(next_states).max(1)[0].detach()
        target_q = rewards + (self.gamma * next_max_q * (1 - dones))

        td_errors = current_q - target_q
        weighted_loss = (td_errors.pow(2) * weights).mean()

        self.optimizer.zero_grad()
        weighted_loss.backward()
        torch.nn.utils.clip_grad_norm_(self.model.parameters(), 1.0)
        self.optimizer.step()

        new_priorities = abs(td_errors.detach().cpu().numpy())
        self.replay_buffer.update_priorities(batch_indices, new_priorities)

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

        return weighted_loss.item()

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())

# === 训练函数 ===
def generate_and_train_for_individual_paths(path_documents, repeats=5, batch_size=32, run_id=1):
    state_dim = 3
    action_dim = 30

    shared_replay_buffer = SharedExperienceReplay(capacity=20000)
    agent = DQNAgentWithPER(state_dim, action_dim, shared_replay_buffer)

    total_cumulative_reward = 0
    path_rewards = {}

    print(f"\n=== 第 {run_id}/20 次训练开始 ===")
    start_time = time.time()

    SAMPLES_PER_BATCH = 50
    NUM_BATCHES = 4
    STEPS_PER_SAMPLE = 3

    for path_idx in range(len(target_paths)):
        path_id = path_idx + 1
        print(f"\n{'=' * 60}")
        print(f"路径 {path_id}/{NUM_PATHS}")
        print(f"{'=' * 60}")

        file_path = os.path.join(path_documents, f"path{path_id}_individual.txt")
        if not os.path.exists(file_path):
            print(f"  文件不存在: {file_path}")
            continue

        path_data = load_path_data(file_path)
        if not path_data:
            print(f"  路径 {path_id} 无有效数据")
            continue

        target_path = target_paths[path_idx]

        if path_idx not in path_rewards:
            path_rewards[path_idx] = 0

        for repeat_idx in range(repeats):
            print(f"\n  重复 {repeat_idx + 1}/{repeats}")

            for batch_idx in range(NUM_BATCHES):
                batch_start = batch_idx * SAMPLES_PER_BATCH
                batch_end = min(batch_start + SAMPLES_PER_BATCH, len(path_data))

                print(f"     批次 {batch_idx + 1}/{NUM_BATCHES} (样本 {batch_start}-{batch_end})")

                for sample_idx in range(batch_start, batch_end):
                    state = path_data[sample_idx]
                    prev_state = None
                    prev_triggered = None

                    for step in range(STEPS_PER_SAMPLE):
                        legal_actions = agent.get_legal_actions(state)

                        if not legal_actions:
                            break

                        state_norm = normalize_state(state)
                        action = agent.act(state_norm, legal_actions)
                        if action is None:
                            break

                        dw, dt, dz = agent.decode_action(action)

                        next_state = (
                            int(np.clip(state[0] + dw, MIN_X, MAX_X)),
                            int(np.clip(state[1] + dt, MIN_Y, MAX_Y)),
                            int(np.clip(state[2] + dz, MIN_Z, MAX_Z))
                        )

                        triggered = section8_hybrid_speed_torque_current(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered,
                                                prev_triggered, prev_state)
                        done = (step == STEPS_PER_SAMPLE - 1)

                        agent.store_transition(state, action, reward, next_state, done)

                        prev_state = state
                        prev_triggered = triggered
                        state = next_state

                        total_cumulative_reward += reward
                        path_rewards[path_idx] += reward

                if len(agent.replay_buffer) >= batch_size:
                    loss = agent.train(batch_size)
                    print(f"         批次 {batch_idx + 1} 完成，损失: {loss:.4f}")

                if (batch_idx + 1) % 2 == 0:
                    agent.update_target_model()
                    print(f"         目标网络已更新 (批次 {batch_idx + 1})")

        print(f"\n路径 {path_id} 完成，累计奖励: {path_rewards[path_idx]:.2f}")
        print(f"共享经验池大小: {len(shared_replay_buffer)}")

    training_time = time.time() - start_time
    print(f"\n=== 第 {run_id}/20 次训练完成，用时: {training_time:.2f} 秒 ===")

    return agent, shared_replay_buffer, total_cumulative_reward, path_rewards, training_time

# === Excel 报告生成 ===
def generate_excel_report(all_runs, similar_group, isolated_group, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    sim_paths = [i+1 for i in similar_group]
    iso_paths = [i+1 for i in isolated_group]

    wb = Workbook()
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    h_color = "4472C4"
    sim_color = "E2EFDA"
    iso_color = "FCE4D6"
    s_color = "FFF2CC"

    # Sheet1: 路径相似度
    ws1 = wb.active
    ws1.title = "路径相似度"
    headers1 = ['Path ID', '分组'] + [f'Run {i}' for i in range(1, 21)] + ['平均', '最高', '最低', '标准差']
    for c, h in enumerate(headers1, 1):
        cell = ws1.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=h_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for pid in range(1, NUM_PATHS + 1):
        row = pid + 1
        if pid in sim_paths:
            gtype, gcolor = "高相似组", sim_color
        elif pid in iso_paths:
            gtype, gcolor = "低相似组", iso_color
        else:
            gtype, gcolor = "未分组", "FFFFFF"

        ws1.cell(row, 1, f"Path {pid}").font = Font(bold=True)
        ws1.cell(row, 2, gtype)
        for c in [1, 2]:
            ws1.cell(row, c).fill = PatternFill("solid", fgColor=gcolor)
            ws1.cell(row, c).alignment = Alignment(horizontal="center")
            ws1.cell(row, c).border = border

        sims = []
        for ri, run in enumerate(all_runs):
            s = run['path_sims'].get(pid, {}).get('avg', 0.0)
            sims.append(s)
            cell = ws1.cell(row, 3 + ri, round(s, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        stats = [np.mean(sims), np.max(sims), np.min(sims), np.std(sims)]
        for i, v in enumerate(stats):
            cell = ws1.cell(row, 23 + i, round(v, 4))
            cell.number_format = '0.0000'
            cell.fill = PatternFill("solid", fgColor=s_color)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    # Sheet2: 分组统计
    ws2 = wb.create_sheet("分组统计")
    headers2 = ['分组', '包含路径'] + [f'Run {i}' for i in range(1, 21)] + ['平均相似度', '标准差']
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=h_color)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    def write_group_row(row, name, paths, color):
        ws2.cell(row, 1, name).font = Font(bold=True)
        ws2.cell(row, 2, ','.join(map(str, paths)))
        for c in [1, 2]:
            ws2.cell(row, c).fill = PatternFill("solid", fgColor=color)
            ws2.cell(row, c).alignment = Alignment(horizontal="center")
            ws2.cell(row, c).border = border
        vals = []
        for ri, run in enumerate(all_runs):
            v = np.mean([run['path_sims'].get(p, {}).get('avg', 0.0) for p in paths])
            vals.append(v)
            cell = ws2.cell(row, 3 + ri, round(v, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        cell = ws2.cell(row, 23, round(np.mean(vals), 4))
        cell.number_format = '0.0000'
        cell.fill = PatternFill("solid", fgColor=s_color)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        cell = ws2.cell(row, 24, round(np.std(vals), 4))
        cell.number_format = '0.0000'
        cell.fill = PatternFill("solid", fgColor=s_color)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    write_group_row(2, "高相似组", sim_paths, sim_color)
    write_group_row(3, "低相似组", iso_paths, iso_color)

    # Sheet3: 轮次汇总
    ws3 = wb.create_sheet("轮次汇总")
    headers3 = ['轮次', '耗时(秒)', '总体平均相似度', '最高相似度', '最低相似度', '高相似组平均', '低相似组平均', '回放池容量']
    for c, h in enumerate(headers3, 1):
        cell = ws3.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=h_color)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for ri, run in enumerate(all_runs, 1):
        row = ri + 1
        high_avg = np.mean([run['path_sims'].get(p, {}).get('avg', 0.0) for p in sim_paths])
        low_avg = np.mean([run['path_sims'].get(p, {}).get('avg', 0.0) for p in iso_paths]) if iso_paths else 0.0
        vals = [
            f"Run {ri}",
            round(run['time'], 2),
            round(run['overall_avg'], 4),
            round(run['max_sim'], 4),
            round(run['min_sim'], 4),
            round(high_avg, 4),
            round(low_avg, 4),
            20000
        ]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row, c, v)
            if c == 1:
                cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    # Sheet4: Top样本
    ws4 = wb.create_sheet("Top样本详情")
    headers4 = ['轮次', '路径', '序号', 'X', 'Y', 'Z', '相似度', '触发规则']
    for c, h in enumerate(headers4, 1):
        cell = ws4.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=h_color)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    r_idx = 2
    for ri, run in enumerate(all_runs, 1):
        for pid in range(1, NUM_PATHS + 1):
            samples = run['samples'].get(pid, [])
            pcolor = sim_color if pid in sim_paths else (iso_color if pid in iso_paths else "FFFFFF")
            for si, (st, _, sim, trig) in enumerate(samples, 1):
                x, y, z = st
                ws4.cell(r_idx, 1, f"Run {ri}").fill = PatternFill("solid", fgColor=pcolor)
                ws4.cell(r_idx, 2, f"Path {pid}").fill = PatternFill("solid", fgColor=pcolor)
                ws4.cell(r_idx, 3, si)
                ws4.cell(r_idx, 4, x)
                ws4.cell(r_idx, 5, y)
                ws4.cell(r_idx, 6, z)
                ws4.cell(r_idx, 7, round(sim, 4)).number_format = '0.0000'
                ws4.cell(r_idx, 8, ','.join(map(str, sorted(trig))))
                for c in range(1, 9):
                    ws4.cell(r_idx, c).alignment = Alignment(horizontal="center")
                    ws4.cell(r_idx, c).border = border
                r_idx += 1

    out_path = os.path.join(out_dir, "20轮训练汇总报告.xlsx")
    wb.save(out_path)
    print(f"Excel 报告已保存: {out_path}")

# === 20 轮总入口 ===
def run_20_times_training():
    model_dir = r"D:\Experiment\CNN\DQNNEW\saved_models_new_vars"
    sample_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_individual"
    report_dir = r"D:\Experiment\CNN\ComparisonExperiment2\excel_reports_new_vars"

    os.makedirs(model_dir, exist_ok=True)
    os.makedirs(sample_dir, exist_ok=True)
    os.makedirs(report_dir, exist_ok=True)

    similar_group, isolated_group = group_paths_by_similarity(target_paths)

    print("=" * 60)
    print("DQN 路径覆盖优化 - 20轮完整训练")
    print(f"参数范围: X [{MIN_X},{MAX_X}] | Y [{MIN_Y},{MAX_Y}] | Z [{MIN_Z},{MAX_Z}]")
    print(f"路径总数: {NUM_PATHS}")
    print(f"高相似组: {[i+1 for i in similar_group]}")
    print(f"低相似组: {[i+1 for i in isolated_group]}")
    print("=" * 60)

    all_runs = []

    for run_id in range(1, 21):
        print(f"\n===== 第 {run_id}/20 轮 =====")
        print("[1/2] 生成初始样本...")
        generate_samples_for_all_paths(num_candidates=2000, top_k=200, run_id=run_id)

        print("[2/2] 开始训练...")
        agent, buffer, total_r, path_r, elapsed = generate_and_train_for_individual_paths(
            sample_dir, repeats=5, batch_size=32, run_id=run_id
        )

        # 保存模型
        model_path = os.path.join(model_dir, f"model_run_{run_id}.pth")
        torch.save({
            'model_state_dict': agent.model.state_dict(),
            'optimizer_state_dict': agent.optimizer.state_dict(),
            'epsilon': agent.epsilon,
            'ranges': {'x': [MIN_X, MAX_X], 'y': [MIN_Y, MAX_Y], 'z': [MIN_Z, MAX_Z]}
        }, model_path)

        # 收集统计
        run_data = {'path_sims': {}, 'samples': {}, 'time': elapsed}
        all_sims = []
        for pi in range(NUM_PATHS):
            pid = pi + 1
            samples = buffer.get_high_reward_samples(target_paths[pi], 20)
            if samples:
                sims = [s[2] for s in samples]
                run_data['path_sims'][pid] = {
                    'avg': np.mean(sims),
                    'max': np.max(sims),
                    'min': np.min(sims)
                }
                run_data['samples'][pid] = samples
                all_sims.extend(sims)
            else:
                run_data['path_sims'][pid] = {'avg': 0.0, 'max': 0.0, 'min': 0.0}
                run_data['samples'][pid] = []

        run_data['overall_avg'] = np.mean(all_sims) if all_sims else 0.0
        run_data['max_sim'] = np.max(all_sims) if all_sims else 0.0
        run_data['min_sim'] = np.min(all_sims) if all_sims else 0.0
        all_runs.append(run_data)

        print(f"本轮完成 | 耗时: {elapsed:.1f}s | 总体平均相似度: {run_data['overall_avg']:.4f}")

    print("\n生成最终 Excel 报告...")
    generate_excel_report(all_runs, similar_group, isolated_group, report_dir)
    print("\n全部 20 轮训练完成！")

if __name__ == "__main__":
    run_20_times_training()