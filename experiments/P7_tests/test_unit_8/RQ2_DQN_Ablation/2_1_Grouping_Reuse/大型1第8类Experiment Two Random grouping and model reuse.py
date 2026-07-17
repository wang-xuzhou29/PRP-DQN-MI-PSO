import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import math

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === 新范围：X:1000~2000, Y:100~200, Z:10~30（整数） ===
STATE_RANGES = {
    'dx': (1000, 2000),
    'dy': (100, 200),
    'dz': (10, 30),
}
STATE_NAMES = ('dx', 'dy', 'dz')
STATE_MIN = np.array([STATE_RANGES[name][0] for name in STATE_NAMES], dtype=np.int32)
STATE_MAX = np.array([STATE_RANGES[name][1] for name in STATE_NAMES], dtype=np.int32)


def clip_state(state):
    """ dx/dy/dz ,  int tuple."""
    return tuple(np.clip(np.array(state, dtype=np.int32), STATE_MIN, STATE_MAX).astype(int))


def random_state():
    """ dx/dy/dz ."""
    return tuple(random.randint(STATE_RANGES[name][0], STATE_RANGES[name][1]) for name in STATE_NAMES)


class StateNormalizer:
    """:  dx/dy/dz  [0, 1]."""

    def __init__(self, ranges=None):
        self.ranges = ranges or STATE_RANGES
        self.names = STATE_NAMES

    def normalize(self, state):
        """ -> ."""
        state = np.array(state, dtype=np.float32)
        normalized = np.zeros_like(state, dtype=np.float32)
        for i, name in enumerate(self.names):
            low, high = self.ranges[name]
            normalized[i] = (state[i] - low) / (high - low)
        return normalized

    def denormalize(self, normalized_state):
        """ ->  dx/dy/dz, ."""
        normalized_state = np.array(normalized_state, dtype=np.float32)
        denormalized = np.zeros_like(normalized_state, dtype=np.float32)
        for i, name in enumerate(self.names):
            low, high = self.ranges[name]
            denormalized[i] = normalized_state[i] * (high - low) + low
            denormalized[i] = np.clip(np.round(denormalized[i]), low, high).astype(int)
        return denormalized


normalizer = StateNormalizer()

def section8_hybrid_speed_torque_current(x, y, z):
    """第8类: 速度扭矩电流混合控制 (速度x, 扭矩y, 电流z) - 144个有效变异分支"""
    triggered = set()  # 记录触发的条件编号（1-144，删除了12个无效分支）
    # 替换原self.standards的固定标准范围（基于代码逻辑推导合理值）
    speed_std = (1000, 2000)
    torque_std = (100, 200)
    current_std = (10, 30)

    # 删除的100%覆盖率分支: [1, 2, 3, 4, 5, 6] - 不包含在函数中
    # 删除的0%覆盖率分支: [39, 40, 59, 66, 74] - 不包含在函数中
    # 删除缺失分支: [137] (原编号148)

    # -------------------------- 1-32：速度扭矩电流基础检查（删除原1-6，保留7-38） --------------------------

    # 编号1 (原编号7)
    if [(x > 1400 and x < 1600 and y > 160 and y < 190)] != [(x > 1400 and x * 8 < 1600 and y > 160 and y < 190)]:
        triggered.add(1)

    # 编号2 (原编号8)
    if [(x > 1400 and x < 1600 and y > 160 and y < 190)] != [(x > 1400 and x < 1600 and y * 10 > 160 and y < 190)]:
        triggered.add(2)

    # 编号3 (原编号9)
    if [(x > 1350 and x < 1650 and z > 11 and z < 14)] != [(x > 1350 and x * 9 < 1650 and z > 11 and z < 14)]:
        triggered.add(3)

    # 编号4 (原编号10)
    if [(x > 1350 and x < 1650 and z > 11 and z < 14)] != [(x > 1350 and x < 1650 and z * 20 > 11 and z < 14)]:
        triggered.add(4)

    # 编号5 (原编号11)
    if [(y > 155 and y < 195 and z > 11 and z < 14)] != [(y > 155 and y * 7 < 195 and z > 11 and z < 14)]:
        triggered.add(5)

    # 编号6 (原编号12)
    if [(y > 155 and y < 195 and z > 11 and z < 14)] != [(y > 155 and y < 195 and z * 22 > 11 and z < 14)]:
        triggered.add(6)

    # 编号7 (原编号13)
    if [(x > 1480 and x < 1520)] != [(x > 1480 and x * 8 < 1520)]:
        triggered.add(7)

    # 编号8 (原编号14)
    if [(x > 1480 and x < 1520)] != [(x * 12 > 1480 and x < 1520)]:
        triggered.add(8)

    # 编号9 (原编号15)
    if [(y > 173 and y < 177)] != [(y * 12 > 173 and y < 177)]:
        triggered.add(9)

    # 编号10 (原编号16)
    if [(y > 173 and y < 177)] != [(y > 173 and y * 22 < 177)]:
        triggered.add(10)

    # 编号11 (原编号17)
    if [(z > 12.2 and z < 12.8)] != [(z * 78 > 12.2 and z < 12.8)]:
        triggered.add(11)

    # 编号12 (原编号18)
    if [(z > 12.2 and z < 12.8)] != [(66 > 12.2 and z < 12.8)]:
        triggered.add(12)

    # 编号13 (原编号19)
    if [(abs(x - 1500) < 40)] != [(abs(x - 1500) < 140)]:
        triggered.add(13)

    # 编号14 (原编号20)
    if [(abs(x - 1500) < 40)] != [(abs(x - 1500) < 990)]:
        triggered.add(14)

    # 编号15 (原编号21)
    if [(abs(y - 175) < 6)] != [(abs(y - 175) < 16)]:
        triggered.add(15)

    # 编号16 (原编号22)
    if [(abs(y - 175) < 6)] != [(abs(y - 175) < 226)]:
        triggered.add(16)

    # 编号17 (原编号23)
    if [(abs(z - 12.5) < 0.4)] != [(abs(z - 12.5) < 4)]:
        triggered.add(17)

    # 编号18 (原编号24)
    if [(abs(z - 12.5) < 0.4)] != [(abs(z - 12.5) < 1)]:
        triggered.add(18)

    # 编号19 (原编号25)
    if [(x > 1350 and y > 155 and z > 11)] != [(x > 1350 and y * 78 > 155 and z > 11)]:
        triggered.add(19)

    # 编号20 (原编号26)
    if [(x > 1350 and y > 155 and z > 11)] != [(x * 67 > 1350 and y > 155 and z > 11)]:
        triggered.add(20)

    # 编号21 (原编号27)
    if [(x < 1650 and y < 195 and z < 14)] != [(x < 1650 and y * 88 < 195 and z < 14)]:
        triggered.add(21)

    # 编号22 (原编号28)
    if [(x < 1650 and y < 195 and z < 14)] != [(x < 1650 and y < 195 and z * 8 < 14)]:
        triggered.add(22)

    # 编号23 (原编号29)
    if [((x / 10 + y + z * 10) > 350 and (x / 10 + y + z * 10) < 410)] != [
        ((x / 10 + y + z * 10) > 3150 and (x / 10 + y + z * 10) < 410)]:
        triggered.add(23)

    # 编号24 (原编号30)
    if [((x / 10 + y + z * 10) > 350 and (x / 10 + y + z * 10) < 410)] != [
        ((x / 10 + y + z * 10) > 350 and (x / 10 + y + z * 10) < 1410)]:
        triggered.add(24)

    # 编号25 (原编号31)
    if [(x * y / 1000 > 230 and x * y / 1000 < 310)] != [(x * y / 1000 > 1230 and x * y / 1000 < 310)]:
        triggered.add(25)

    # 编号26 (原编号32)
    if [(x * y / 1000 > 230 and x * y / 1000 < 310)] != [(x * y / 1000 > 230 and x * y / 1000 < 3110)]:
        triggered.add(26)

    # 编号27 (原编号33)
    if [(x * z / 100 > 160 and x * z / 100 < 220)] != [(x * z / 100 > 1610 and x * z / 100 < 220)]:
        triggered.add(27)

    # 编号28 (原编号34)
    if [(x * z / 100 > 160 and x * z / 100 < 220)] != [(x * z / 100 > 160 and x * z / 100 < 2120)]:
        triggered.add(28)

    # 编号29 (原编号35)
    if [(y * z > 2000 and y * z < 2600)] != [(y * z > 200 and y * z < 2600)]:
        triggered.add(29)

    # 编号30 (原编号36)
    if [(y * z > 2000 and y * z < 2600)] != [(y * z > 2000 and y * z < 260)]:
        triggered.add(30)

    # 编号31 (原编号37)
    if [((x / 10 + y + z * 10) / 3 > 116 and (x / 10 + y + z * 10) / 3 < 136)] != [
        ((x / 10 + y + z * 10) / 5 > 116 and (x / 10 + y + z * 10) / 3 < 136)]:
        triggered.add(31)

    # 编号32 (原编号38)
    if [((x / 10 + y + z * 10) / 3 > 116 and (x / 10 + y + z * 10) / 3 < 136)] != [
        ((x / 10 + y + z * 10) / 3 > 116 and (x / 10 + y + z * 10) / 11 < 136)]:
        triggered.add(32)

    # 跳过原编号39, 40 (0%覆盖率分支)

    # -------------------------- 33-58：驱动系统分析（原41-80，删除原39、40） --------------------------

    # 编号33 (原编号41)
    if [(x / (y * 10) > 0.8 and x / (y * 10) < 1.0)] != [(x / (y * 10) > 8 and x / (y * 10) < 1.0)]:
        triggered.add(33)

    # 编号34 (原编号42)
    if [(x / (y * 10) > 0.8 and x / (y * 10) < 1.0)] != [(x / (y * 10) > 0.8 and x / (y * 10) < 10)]:
        triggered.add(34)

    # 编号35 (原编号43)
    if [(x / z > 110 and x / z < 140)] != [(x / z > 1220 and x / z < 140)]:
        triggered.add(35)

    # 编号36 (原编号44)
    if [(x / z > 110 and x / z < 140)] != [(x / z > 110 and x / z < 14)]:
        triggered.add(36)

    # 编号37 (原编号45)
    if [(y / z > 12 and y / z < 16)] != [(y / z > 112 and y / z < 16)]:
        triggered.add(37)

    # 编号38 (原编号46)
    if [(y / z > 12 and y / z < 16)] != [(y / z > 12 and y / z < 1226)]:
        triggered.add(38)

    # 编号39 (原编号47)
    if [((x - 1200) / 10 + (y - 150) > 35 and (x - 1200) / 10 + (y - 150) < 55)] != [
        ((x - 1200) / 10 + (y - 150) > 315 and (x - 1200) / 10 + (y - 150) < 55)]:
        triggered.add(39)

    # 编号40 (原编号48)
    if [((x - 1200) / 10 + (y - 150) > 35 and (x - 1200) / 10 + (y - 150) < 55)] != [
        ((x - 1200) / 10 + (y - 220) > 35 and (x - 1200) / 10 + (y - 150) < 55)]:
        triggered.add(40)

    # 编号41 (原编号49)
    if [((x - 1200) / 10 + (z - 10) * 5 > 40 and (x - 1200) / 10 + (z - 10) * 5 < 60)] != [
        ((x - 1200) / 10 + (z - 10) * 15 > 40 and (x - 1200) / 10 + (z - 10) * 5 < 60)]:
        triggered.add(41)

    # 编号42 (原编号50)
    if [((x - 1200) / 10 + (z - 10) * 5 > 40 and (x - 1200) / 10 + (z - 10) * 5 < 60)] != [
        ((x - 1200) / 10 + (z - 10) * 115 > 40 and (x - 1200) / 10 + (z - 10) * 5 < 60)]:
        triggered.add(42)

    # 编号43 (原编号51)
    if [((y - 150) + (z - 10) * 2 > 30 and (y - 150) + (z - 10) * 2 < 50)] != [
        ((y - 150) + (z - 10) * 2 > 302 and (y - 150) + (z - 10) * 2 < 50)]:
        triggered.add(43)

    # 编号44 (原编号52)
    if [((y - 150) + (z - 10) * 2 > 30 and (y - 150) + (z - 10) * 2 < 50)] != [
        ((y - 150) + (z - 10) * 211 > 30 and (y - 150) + (z - 10) * 2 < 50)]:
        triggered.add(44)

    # 编号45 (原编号53)
    if [(abs((x - 1500) / 10 - (y - 175)) < 12)] != [(abs((x - 1500) / 10 - (y - 175)) < 112)]:
        triggered.add(45)

    # 编号46 (原编号54)
    if [(abs((x - 1500) / 10 - (y - 175)) < 12)] != [(abs((x - 1500) / 101 - (y - 175)) < 12)]:
        triggered.add(46)

    # 编号47 (原编号55)
    if [(abs((x - 1500) / 100 - (z - 12.5)) < 3)] != [(abs((x - 1500) / 100 - (z - 12.5)) < 43)]:
        triggered.add(47)

    # 编号48 (原编号56)
    if [(abs((x - 1500) / 100 - (z - 12.5)) < 3)] != [(abs((x - 1500) / 100 - (z - 12.5)) < 13)]:
        triggered.add(48)

    # 编号49 (原编号57)
    if [(abs((y - 175) - (z - 12.5) * 12) < 8)] != [(abs((y - 175) - (z - 12.5) * 12) < 118)]:
        triggered.add(49)

    # 编号50 (原编号58)
    if [(abs((y - 175) - (z - 12.5) * 12) < 8)] != [(abs((y - 175) - (z - 12.5) * 232) < 8)]:
        triggered.add(50)

    # 跳过原编号59 (0%覆盖率分支)

    # 编号51 (原编号60)
    if [(x / (y * 10 + 500) > 0.7 and x / (y * 10 + 500) < 0.9)] != [
        (x / (y * 10 + 500) > 337 and x / (y * 10 + 500) < 0.9)]:
        triggered.add(51)

    # 编号52 (原编号61)
    if [(y / (z + 5) > 9 and y / (z + 5) < 13)] != [(y / (z + 5) > 119 and y / (z + 5) < 13)]:
        triggered.add(52)

    # 编号53 (原编号62)
    if [(y / (z + 5) > 9 and y / (z + 5) < 13)] != [(y / (z + 5) > 9 and y / (z + 5) < 123)]:
        triggered.add(53)

    # 编号54 (原编号63)
    if [(z / (x / 150) > 1.1 and z / (x / 150) < 1.5)] != [(z / (x / 150) > 11 and z / (x / 150) < 1.5)]:
        triggered.add(54)

    # 编号55 (原编号64)
    if [(z / (x / 150) > 1.1 and z / (x / 150) < 1.5)] != [(z / (x / 150) > 1.1 and z / (x / 150) < 15)]:
        triggered.add(55)

    # 编号56 (原编号65)
    if [((x * y * z / 10000) > 28 and (x * y * z / 10000) < 38)] != [
        ((x * y * z / 10000) > 28 and (x * y * z / 10000) < 238)]:
        triggered.add(56)

    # 跳过原编号66 (0%覆盖率分支)

    # 编号57 (原编号67)
    if [(x / 20 + y * 0.6 + z * 8 > 260 and x / 20 + y * 0.6 + z * 8 < 300)] != [
        (x / 20 + y * 0.6 + z * 10 > 260 and x / 20 + y * 0.6 + z * 8 < 300)]:
        triggered.add(57)

    # 编号58 (原编号68)
    if [(x / 20 + y * 0.6 + z * 8 > 260 and x / 20 + y * 0.6 + z * 8 < 300)] != [
        (x / 20 + y * 0.6 + z * 8 > 260 and x / 20 + y * 6 + z * 8 < 300)]:
        triggered.add(58)

    # 编号59 (原编号69)
    if [((x / 1500) ** 0.4 * (y / 175) ** 0.4 * (z / 12.5) ** 0.2 > 0.93)] != [
        ((x / 1500) ** 0.4 * (y / 175) ** 0.4 * (z / 12.5) ** 0.2 > 93)]:
        triggered.add(59)

    # 编号60 (原编号70)
    if [((x / 1500) ** 0.4 * (y / 175) ** 0.4 * (z / 12.5) ** 0.2 > 0.93)] != [
        ((x / 1500) ** 0.4 * (y / 175) ** 4 * (z / 12.5) ** 0.2 > 0.93)]:
        triggered.add(60)

    # 编号61 (原编号71)
    if [((x - 1500) / 10 * (y - 175) > -200 and (x - 1500) / 10 * (y - 175) < 200)] != [
        ((x - 1500) / 110 * (y - 175) > -200 and (x - 1500) / 10 * (y - 175) < 200)]:
        triggered.add(61)

    # 编号62 (原编号72)
    if [((x - 1500) / 10 * (y - 175) > -200 and (x - 1500) / 10 * (y - 175) < 200)] != [
        ((x - 1500) / 10 * (y - 175) > -200 and (x - 1500) / 110 * (y - 175) < 200)]:
        triggered.add(62)

    # 编号63 (原编号73)
    if [((x - 1500) / 100 * (z - 12.5) > -20 and (x - 1500) / 100 * (z - 12.5) < 20)] != [
        ((x - 1500) / 100 * (z - 12.5) > -20 and (x - 1500) / 10 * (z - 12.5) < 20)]:
        triggered.add(63)

    # 跳过原编号74 (0%覆盖率分支)

    # 编号64 (原编号75)
    if [((y - 175) * (z - 12.5) > -20 and (y - 175) * (z - 12.5) < 20)] != [
        ((y - 175) * (z - 12.5) > -20 and (y - 175) * (z - 12.5) < 210)]:
        triggered.add(64)

    # 编号65 (原编号76)
    if [((y - 175) * (z - 12.5) > -20 and (y - 175) * (z - 12.5) < 20)] != [((y - 175) * (z - 125) < 20)]:
        triggered.add(65)

    # 编号66 (原编号77)
    if [(abs(x / 1500 + y / 175 + z / 12.5 - 3) < 0.2)] != [(abs(x / 1500 + y / 175 + z / 12.5 - 3) < 112)]:
        triggered.add(66)

    # 编号67 (原编号78)
    if [(abs(x / 1500 + y / 175 + z / 12.5 - 3) < 0.2)] != [(abs(x / 1500 + y / 175 + z / 12.5 - 3) < 2222)]:
        triggered.add(67)

    # 编号68 (原编号79)
    if [(abs((x / 1500) * (y / 175) * (z / 12.5) - 1) < 0.15)] != [
        (abs((x / 1500) * (y / 175) * (z / 12.5) - 1) < 115)]:
        triggered.add(68)

    # 编号69 (原编号80)
    if [(abs((x / 1500) * (y / 175) * (z / 12.5) - 1) < 0.15)] != [
        (abs((x / 1500) * (y / 175) * (z / 12.5) - 1) < 2225)]:
        triggered.add(69)

    # -------------------------- 70-107：动力传动协调（原81-120） --------------------------

    # 编号70 (原编号81)
    if [(x > 1480 and x < 1520 and y > 172 and y < 178 and z > 12.2 and z < 12.8)] != [
        (x > 1480 and x < 1520 and y > 172 and y < 178 and z > 12.2 and 8 < 12.8)]:
        triggered.add(70)

    # 编号71 (原编号82)
    if [(x > 1480 and x < 1520 and y > 172 and y < 178 and z > 12.2 and z < 12.8)] != [
        (x > 1480 and x < 1520 and y > 172 and y < 178 and z > 12.2 and 10 < 12.8)]:
        triggered.add(71)

    # 编号72 (原编号83)
    if [(abs(x / (y * 10) - 0.86) < 0.06 and abs(z - 12.5) < 0.3)] != [
        (abs(x / (y * 10) - 0.86) < 0.06 and abs(z - 12.5) < 3)]:
        triggered.add(72)

    # 编号73 (原编号84)
    if [(abs(x / (y * 10) - 0.86) < 0.06 and abs(z - 12.5) < 0.3)] != [
        (abs(x / (y * 10) - 0.86) < 0.06 and abs(z - 12.5) < 13)]:
        triggered.add(73)

    # 编号74 (原编号85)
    if [(x * y / 1000 > 258 and x * y / 1000 < 268 and z > 12.3 and z < 12.7)] != [
        (x * y / 1000 > 258 and x * y / 1000 < 268 and 8 * z > 12.3 and z < 12.7)]:
        triggered.add(74)

    # 编号75 (原编号86)
    if [(x * y / 1000 > 258 and x * y / 1000 < 268 and z > 12.3 and z < 12.7)] != [
        (x * y / 1000 > 258 and x * y / 1000 < 268 and z * 99 > 12.3 and z < 12.7)]:
        triggered.add(75)

    # 编号76 (原编号87)
    if [((x / 10 + y) / 2 > 137 and (x / 10 + y) / 2 < 143 and z > 12.2 and z < 12.8)] != [
        ((x / 10 + y) / 2 > 137 and (x / 10 + y) / 2 < 143 and z > 12.2 and 8 < 12.8)]:
        triggered.add(76)

    # 编号77 (原编号88)
    if [((x / 10 + y) / 2 > 137 and (x / 10 + y) / 2 < 143 and z > 12.2 and z < 12.8)] != [
        ((x / 10 + y) / 2 > 137 and (x / 10 + y) / 2 < 143 and z > 12.2 and 10 < 12.8)]:
        triggered.add(77)

    # 编号78 (原编号89)
    if [(abs(x / 10 - y) < 25 and z > 12.2 and z < 12.8)] != [(abs(x / 10 - y) < 25 and z * 12 > 12.2 and z < 12.8)]:
        triggered.add(78)

    # 编号79 (原编号90)
    if [(abs(x / 10 - y) < 25 and z > 12.2 and z < 12.8)] != [(abs(x / 10 - y) < 25 and z * 89 > 12.2 and z < 12.8)]:
        triggered.add(79)

    # 编号80 (原编号91)
    if [(math.sqrt((x / 10 - 150) ** 2 + (y - 175) ** 2) < 6 and abs(z - 12.5) < 0.3)] != [
        (math.sqrt((x / 10 - 150) ** 2 + (y - 175) ** 2) < 6 and abs(z - 12.5) < 3)]:
        triggered.add(80)

    # 编号81 (原编号92)
    if [(math.sqrt((x / 10 - 150) ** 2 + (y - 175) ** 2) < 6 and abs(z - 12.5) < 0.3)] != [
        (math.sqrt((x / 10 - 150) ** 2 + (y - 175) ** 2) < 6 and abs(z - 12.5) < 13)]:
        triggered.add(81)

    # 编号82 (原编号93)
    if [(x / 1500 > 0.98 and x / 1500 < 1.02 and y / 175 > 0.97 and y / 175 < 1.03)] != [
        (x / 1500 > 0.98 and x / 1500 < 1.02 and y / 175 > 97 and y / 175 < 1.03)]:
        triggered.add(82)

    # 编号83 (原编号94)
    if [(x / 1500 > 0.98 and x / 1500 < 1.02 and y / 175 > 0.97 and y / 175 < 1.03)] != [
        (x / 1500 > 0.98 and x / 1500 < 1.02 and y * 8 / 175 > 0.97 and y / 175 < 1.03)]:
        triggered.add(83)

    # 编号84 (原编号95)
    if [(z / 12.5 > 0.976 and z / 12.5 < 1.024)] != [(z / 12.5 > 0.976 and z / 125 < 1.024)]:
        triggered.add(84)

    # 编号85 (原编号96)
    if [(z / 12.5 > 0.976 and z / 12.5 < 1.024)] != [(z / 12.5 > 0.976 and z / 200 < 1.024)]:
        triggered.add(85)

    # 编号86 (原编号97)
    if [((x / 1500 + y / 175 + z / 12.5) / 3 > 0.98 and (x / 1500 + y / 175 + z / 12.5) / 3 < 1.02)] != [
        ((x / 1500 + y / 175 + z / 12.5) / 3 > 98 and (x / 1500 + y / 175 + z / 12.5) / 3 < 1.02)]:
        triggered.add(86)

    # 编号87 (原编号98)
    if [((x / 1500 + y / 175 + z / 12.5) / 3 > 0.98 and (x / 1500 + y / 175 + z / 12.5) / 3 < 1.02)] != [
        ((x / 1500 + y / 175 + z / 12.5) / 23 > 0.98 and (x / 1500 + y / 175 + z / 12.5) / 3 < 1.02)]:
        triggered.add(87)

    # 编号88 (原编号99)
    if [(max(abs(x / 1500 - 1), abs(y / 175 - 1), abs(z / 12.5 - 1)) < 0.025)] != [
        (max(abs(x / 1500 - 1), abs(y / 175 - 1), abs(z / 12.5 - 1)) < 0.55)]:
        triggered.add(88)

    # 编号89 (原编号100)
    if [(max(abs(x / 1500 - 1), abs(y / 175 - 1), abs(z / 12.5 - 1)) < 0.025)] != [
        (max(abs(x / 1500 - 1), abs(y / 175 - 1), abs(z / 12.5 - 1)) < 0.25)]:
        triggered.add(89)

    # 编号90 (原编号101)
    if [(min(x / 1500, y / 175, z / 12.5) > 0.98)] != [(min(x / 1500, y / 175, z / 12.5) > 498)]:
        triggered.add(90)

    # 编号91 (原编号102)
    if [(min(x / 1500, y / 175, z / 12.5) > 0.98)] != [(min(x / 1500, y / 1375, z / 12.5) > 0.98)]:
        triggered.add(91)

    # 编号92 (原编号103)
    if [(max(x / 1500, y / 175, z / 12.5) < 1.02)] != [(max(x / 1500, y / 175, z / 12.5) < 302)]:
        triggered.add(92)

    # 编号93 (原编号104)
    if [(max(x / 1500, y / 175, z / 12.5) < 1.02)] != [(max(x / 1500, y / 175, z / 12.5) < 102)]:
        triggered.add(93)

    # 编号94 (原编号105)
    if [(abs(max(x / 10, y, z * 10) - min(x / 10, y, z * 10)) < 30)] != [
        (abs(max(x / 10, y, z * 120) - min(x / 10, y, z * 10)) < 30)]:
        triggered.add(94)

    # 编号95 (原编号106)
    if [(abs(max(x / 10, y, z * 10) - min(x / 10, y, z * 10)) < 30)] != [
        (abs(max(x / 10, y, z * 10) - min(x / 10, y, z * 10)) < 230)]:
        triggered.add(95)

    # 编号96 (原编号107)
    if [((x / 10 + y + z * 10) > 378 and (x / 10 + y + z * 10) < 382)] != [
        ((x / 10 + y + z * 10) > 378 and (x / 10 + y + z * 10) < 3112)]:
        triggered.add(96)

    # 编号97 (原编号108)
    if [((x / 10 + y + z * 10) > 378 and (x / 10 + y + z * 10) < 382)] != [
        ((x / 10 + y + z * 10) > 3781 and (x / 10 + y + z * 10) < 382)]:
        triggered.add(97)

    # 编号98 (原编号109)
    if [((x * y * z / 10000) > 32.5 and (x * y * z / 10000) < 33.5)] != [
        ((x * y * z / 10000) > 32.5 and (x * y * 2 / 10000) < 33.5)]:
        triggered.add(98)

    # 编号99 (原编号110)
    if [((x * y * z / 10000) > 32.5 and (x * y * z / 10000) < 33.5)] != [
        ((x * y * z / 10000) > 32.5 and (x * y * 2.4 / 10000) < 33.5)]:
        triggered.add(99)

    # 编号100 (原编号111)
    if [(abs((x / 10 + y + z * 10) / 3 - 126.7) < 1.5)] != [(abs((x / 10 + y + z * 10) / 3 - 126.7) < 15)]:
        triggered.add(100)

    # 编号101 (原编号112)
    if [(abs((x / 10 + y + z * 10) / 3 - 126.7) < 1.5)] != [(abs((x / 10 + y + z * 10) / 3 - 126.7) < 115)]:
        triggered.add(101)

    # 编号102 (原编号113)
    if [(x > 1495 and x < 1505 and y > 174 and y < 176)] != [(x > 1495 and x < 1505 and y * 10 > 174 and y < 176)]:
        triggered.add(102)

    # 编号103 (原编号114)
    if [(x > 1495 and x < 1505 and y > 174 and y < 176)] != [(x > 1495 and x < 1505 and y * 88 > 174 and y < 176)]:
        triggered.add(103)

    # 编号104 (原编号115)
    if [(z > 12.45 and z < 12.55)] != [(z * 10 > 12.45 and z < 12.55)]:
        triggered.add(104)

    # 编号105 (原编号116)
    if [(z > 12.45 and z < 12.55)] != [(z * 8 > 12.45 and z < 12.55)]:
        triggered.add(105)

    # -------------------------- 106-144：动力控制优化（原117-156，删除原148） --------------------------

    # 编号106 (原编号117)
    if [(x < 1350 or y < 160 or z < 11.5)] != [(x < 1350 or y * 8 < 160 or z < 11.5)]:
        triggered.add(106)

    # 编号107 (原编号118)
    if [(x < 1350 or y < 160 or z < 11.5)] != [(x < 1350 or y < 160 or z * 89 < 11.5)]:
        triggered.add(107)

    # 编号108 (原编号119)
    if [(x > 1650 or y > 190 or z > 13.5)] != [(x > 1650 or y > 190 or z > 135)]:
        triggered.add(108)

    # 编号109 (原编号120)
    if [(x > 1650 or y > 190 or z > 13.5)] != [(x > 1650 or y * 87 > 190 or z > 13.5)]:
        triggered.add(109)

    # 编号110 (原编号121)
    if [(abs(x / (y * 10) - 0.86) > 0.1)] != [(abs(x / (y * 10) - 0.86) > 111)]:
        triggered.add(110)

    # 编号111 (原编号122)
    if [(abs(x / (y * 10) - 0.86) > 0.1)] != [(abs(x / (y * 10) - 0.86) > 1671)]:
        triggered.add(111)

    # 编号112 (原编号123)
    if [(abs(x / z - 120) > 15)] != [(abs(x / z - 120) > 1522)]:
        triggered.add(112)

    # 编号113 (原编号124)
    if [(abs(x / z - 120) > 15)] != [(abs(x / z - 120) > 115)]:
        triggered.add(113)

    # 编号114 (原编号125)
    if [(abs(y / z - 14) > 2)] != [(abs(y / z - 14) > 112)]:
        triggered.add(114)

    # 编号115 (原编号126)
    if [(abs(y / z - 14) > 2)] != [(abs(y / z - 14) > 211)]:
        triggered.add(115)

    # 编号116 (原编号127)
    if [((x / 10 + y + z * 10) < 360 or (x / 10 + y + z * 10) > 400)] != [
        ((x / 10 + y + z * 10) < 36 or (x / 10 + y + z * 10) > 400)]:
        triggered.add(116)

    # 编号117 (原编号128)
    if [((x / 10 + y + z * 10) < 360 or (x / 10 + y + z * 10) > 400)] != [
        ((x / 10 + y + z * 10) < 989 or (x / 10 + y + z * 10) > 400)]:
        triggered.add(117)

    # 编号118 (原编号129)
    if [((x * y * z / 10000) < 30 or (x * y * z / 10000) > 36)] != [
        ((x * y * z / 10000) < 30 or (x * y * z / 10000) > 136)]:
        triggered.add(118)

    # 编号119 (原编号130)
    if [((x * y * z / 10000) < 30 or (x * y * z / 10000) > 36)] != [
        ((x * y * z / 10000) < 30 or (x * y * z / 10000) > 316)]:
        triggered.add(119)

    # 编号120 (原编号131)
    if [(x < 1300 and y < 160)] != [(x < 1300 and y * 878 < 160)]:
        triggered.add(120)

    # 编号121 (原编号132)
    if [(x < 1300 and y < 160)] != [(x * 71 < 1300 and y < 160)]:
        triggered.add(121)

    # 编号122 (原编号133)
    if [(x > 1700 and y > 190)] != [(x > 1700 and y * 78 > 190)]:
        triggered.add(122)

    # 编号123 (原编号134)
    if [(x > 1700 and y > 190)] != [(x * 78 > 1700 and y > 190)]:
        triggered.add(123)

    # 编号124 (原编号135)
    if [(x < 1300 and z < 11)] != [(x < 1300 and z * 91 < 11)]:
        triggered.add(124)

    # 编号125 (原编号136)
    if [(x < 1300 and z < 11)] != [(x * 12 < 1300 and z < 11)]:
        triggered.add(125)

    # 编号126 (原编号137)
    if [(x > 1700 and z > 14)] != [(x > 1700 and z * 21 > 14)]:
        triggered.add(126)

    # 编号127 (原编号138)
    if [(x > 1700 and z > 14)] != [(x * 123 > 1700 and z > 14)]:
        triggered.add(127)

    # 编号128 (原编号139)
    if [(y < 160 and z < 11)] != [(y < 160 and z * 78 < 11)]:
        triggered.add(128)

    # 编号129 (原编号140)
    if [(y < 160 and z < 11)] != [(y * 8 < 160 and z < 11)]:
        triggered.add(129)

    # 编号130 (原编号141)
    if [(y > 190 and z > 14)] != [(y * 8 > 190 and z > 14)]:
        triggered.add(130)

    # 编号131 (原编号142)
    if [(y > 190 and z > 14)] != [(y > 190 and z * 9 > 14)]:
        triggered.add(131)

    # 编号132 (原编号143)
    if [(x < 1250 or x > 1750)] != [(x * 67 < 1250 or x > 1750)]:
        triggered.add(132)

    # 编号133 (原编号144)
    if [(x < 1250 or x > 1750)] != [(x < 1250 or x * 53 > 1750)]:
        triggered.add(133)

    # 编号134 (原编号145)
    if [(y < 145 or y > 205)] != [(y * 67 < 145 or y > 205)]:
        triggered.add(134)

    # 编号135 (原编号146)
    if [(y < 145 or y > 205)] != [(y < 145 or y * 67 > 205)]:
        triggered.add(135)

    # 编号136 (原编号147)
    if [(z < 9.5 or z > 15.5)] != [(z < 9.5 or z * 67 > 15.5)]:
        triggered.add(136)

    # 跳过原编号148 (缺失分支)

    # 编号137 (原编号149)
    if [(x < 1200 and y < 155 and z < 10.5)] != [(x * 67 < 1200 and y < 155 and z < 10.5)]:
        triggered.add(137)

    # 编号138 (原编号150)
    if [(x < 1200 and y < 155 and z < 10.5)] != [(x < 1200 and y < 55 and z < 10.5)]:
        triggered.add(138)

    # 编号139 (原编号151)
    if [(x > 1800 and y > 195 and z > 14.5)] != [(x > 1800 and y * 78 > 195 and z > 14.5)]:
        triggered.add(139)

    # 编号140 (原编号152)
    if [(x > 1800 and y > 195 and z > 14.5)] != [(x > 1800 and y > 966 and z > 14.5)]:
        triggered.add(140)

    # 编号141 (原编号153)
    if [((x / 10 - 150) ** 2 + (y - 175) ** 2 + (z - 12.5) ** 2 * 100) > 150] != [
        ((x / 10 - 150) ** 6 + (y - 175) ** 2 + (z - 12.5) ** 2 * 100) > 150]:
        triggered.add(141)

    # 编号142 (原编号154)
    if [((x / 10 - 150) ** 2 + (y - 175) ** 2 + (z - 12.5) ** 2 * 100) > 150] != [
        ((x / 10 - 150) ** 2 + (y - 175) ** 8 + (z - 12.5) ** 2 * 100) > 150]:
        triggered.add(142)

    # 编号143 (原编号155)
    if [(abs((x * y * z / 10000) / 33 - 1) > 0.18)] != [(abs((x * y * z / 10000) / 33 - 1) > 18)]:
        triggered.add(143)

    # 编号144 (原编号156)
    if [(abs((x * y * z / 10000) / 33 - 1) > 0.18)] != [(abs((x * y * z / 10000) / 33 - 1) > 10)]:
        triggered.add(144)

    return triggered


# === reward function ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward


# === 将规则函数赋给 execute_Tr （现在指向 section8） ===
def execute_Tr(x, y, z):
    return section8_hybrid_speed_torque_current(x, y, z)


# === 目标路径组（对应 section8 的 144 条规则） ===
targetPaths = [
    {1, 4, 6, 9, 11, 12, 13, 14, 15, 16, 17, 21, 22, 24, 25, 27, 29, 32, 33, 38, 44, 45, 46, 47, 48, 49, 51, 52, 58, 59,
     60, 64, 68, 69, 74, 75, 78, 79, 88, 89, 92, 93, 95, 96, 101, 104, 105, 107, 109, 110, 111, 112, 113, 114, 115, 119,
     133, 135, 136, 143, 144},

    {4, 6, 9, 11, 12, 13, 14, 15, 16, 17, 21, 22, 24, 25, 27, 29, 32, 34, 37, 39, 40, 44, 45, 47, 48, 50, 51, 52, 58,
     59, 60, 61, 64, 68, 69, 74, 75, 78, 79, 88, 89, 92, 93, 95, 96, 101, 104, 105, 109, 110, 111, 112, 113, 119, 128,
     129, 133, 135, 136, 143, 144},

    {3, 5, 9, 11, 12, 13, 14, 16, 17, 18, 21, 22, 24, 25, 27, 29, 32, 34, 35, 36, 37, 39, 40, 44, 45, 49, 51, 52, 54,
     58, 59, 60, 61, 65, 74, 75, 78, 79, 86, 87, 88, 89, 92, 93, 95, 96, 101, 104, 105, 106, 109, 110, 111, 112, 113,
     119, 133, 135, 136, 143, 144},

    {2, 3, 8, 9, 11, 12, 13, 14, 16, 17, 18, 19, 21, 22, 23, 27, 29, 31, 34, 35, 36, 41, 42, 44, 45, 49, 51, 54, 56, 57,
     62, 64, 66, 67, 68, 69, 78, 79, 88, 89, 94, 96, 100, 101, 104, 105, 106, 109, 110, 111, 114, 115, 117, 119, 133,
     134, 136, 143, 144},

    {2, 4, 7, 9, 11, 12, 16, 17, 21, 22, 24, 27, 29, 32, 33, 35, 36, 37, 41, 42, 44, 45, 50, 51, 52, 57, 64, 66, 67, 68,
     69, 78, 79, 83, 88, 89, 95, 96, 100, 101, 102, 103, 104, 105, 109, 110, 111, 112, 113, 119, 128, 129, 133, 135,
     136, 143, 144},

    {6, 8, 11, 12, 14, 15, 16, 17, 21, 22, 24, 29, 32, 35, 36, 38, 39, 40, 42, 43, 45, 49, 52, 54, 56, 57, 60, 61, 63,
     65, 66, 67, 68, 69, 88, 89, 92, 93, 95, 96, 100, 101, 104, 105, 109, 110, 111, 114, 115, 119, 124, 125, 132, 135,
     136, 143, 144},

    {1, 4, 6, 8, 9, 11, 12, 13, 14, 15, 16, 17, 21, 22, 24, 25, 29, 32, 33, 35, 36, 37, 39, 40, 42, 44, 49, 52, 54, 57,
     59, 60, 65, 66, 67, 68, 69, 72, 73, 78, 79, 88, 89, 95, 96, 100, 101, 104, 105, 107, 109, 119, 133, 135, 136, 143,
     144},

    {1, 3, 5, 7, 9, 11, 12, 17, 18, 21, 22, 24, 25, 27, 30, 32, 33, 35, 36, 37, 39, 40, 41, 42, 44, 50, 52, 54, 58, 59,
     65, 72, 73, 74, 75, 78, 79, 80, 81, 82, 88, 89, 95, 96, 101, 102, 103, 104, 105, 109, 119, 133, 135, 136, 142, 143,
     144},

    {8, 9, 11, 12, 14, 16, 17, 21, 22, 23, 29, 31, 33, 35, 36, 37, 44, 46, 49, 54, 56, 62, 63, 64, 66, 67, 68, 69, 72,
     73, 78, 79, 88, 89, 95, 98, 100, 101, 104, 105, 109, 117, 119, 120, 121, 124, 125, 128, 129, 132, 134, 136, 137,
     138, 143, 144},

    {9, 11, 12, 14, 16, 17, 24, 25, 27, 29, 32, 34, 37, 39, 40, 45, 47, 48, 50, 51, 52, 57, 59, 60, 61, 64, 68, 69, 74,
     75, 78, 79, 88, 89, 92, 93, 95, 96, 100, 101, 104, 105, 110, 111, 112, 113, 119, 122, 126, 128, 129, 133, 135, 136,
     143, 144},

    {6, 8, 11, 12, 14, 16, 17, 21, 22, 23, 29, 31, 35, 36, 38, 39, 40, 43, 45, 49, 52, 54, 56, 57, 60, 61, 63, 65, 66,
     67, 68, 69, 88, 89, 92, 93, 95, 96, 100, 101, 104, 105, 110, 111, 114, 115, 119, 123, 124, 125, 131, 132, 135, 136,
     143, 144},

    {8, 9, 11, 12, 14, 16, 17, 21, 22, 29, 34, 35, 36, 42, 44, 45, 49, 51, 54, 56, 62, 63, 64, 66, 67, 68, 69, 78, 79,
     88, 94, 98, 99, 100, 101, 104, 105, 109, 110, 111, 114, 115, 116, 119, 120, 121, 124, 125, 128, 129, 133, 134, 136,
     143, 144},

    {5, 8, 11, 12, 14, 16, 17, 20, 21, 22, 24, 30, 32, 38, 39, 40, 42, 43, 45, 49, 52, 54, 58, 60, 61, 63, 65, 66, 67,
     68, 69, 88, 89, 92, 93, 95, 96, 100, 101, 104, 105, 110, 111, 112, 113, 114, 115, 119, 123, 131, 132, 135, 136,
     143, 144},

    {8, 9, 14, 16, 17, 23, 27, 29, 31, 34, 41, 42, 44, 45, 47, 48, 49, 51, 55, 56, 57, 62, 66, 67, 68, 69, 84, 85, 88,
     92, 93, 95, 96, 98, 99, 100, 101, 108, 110, 111, 112, 113, 114, 115, 117, 119, 120, 121, 127, 130, 132, 134, 136,
     143, 144},

    {5, 8, 9, 14, 15, 16, 17, 18, 20, 21, 22, 23, 30, 32, 37, 41, 42, 44, 45, 46, 47, 48, 49, 52, 55, 58, 62, 65, 66,
     67, 68, 69, 76, 77, 84, 85, 88, 89, 92, 93, 95, 96, 100, 101, 109, 110, 111, 112, 113, 119, 132, 135, 136, 143,
     144},

    {1, 3, 5, 9, 11, 12, 17, 18, 21, 22, 24, 25, 27, 30, 32, 33, 35, 36, 37, 39, 40, 44, 50, 52, 54, 58, 59, 60, 65, 72,
     73, 74, 75, 78, 79, 80, 81, 82, 88, 89, 95, 96, 101, 104, 105, 109, 119, 133, 135, 136, 141, 142, 143, 144},

    {1, 4, 6, 8, 10, 11, 12, 17, 21, 22, 24, 25, 29, 32, 33, 35, 36, 38, 39, 40, 42, 44, 49, 52, 58, 59, 68, 69, 72, 73,
     74, 75, 80, 81, 82, 88, 89, 95, 96, 101, 104, 105, 107, 109, 112, 113, 114, 115, 119, 133, 135, 136, 143, 144},

    {8, 9, 11, 12, 14, 16, 17, 21, 22, 23, 29, 31, 33, 35, 36, 37, 42, 44, 46, 50, 52, 54, 56, 62, 63, 64, 66, 67, 68,
     69, 72, 73, 78, 79, 88, 89, 95, 97, 104, 105, 109, 117, 119, 120, 121, 124, 125, 128, 129, 133, 135, 136, 143,
     144},

    {8, 9, 11, 12, 14, 16, 17, 21, 22, 29, 34, 42, 44, 45, 49, 51, 54, 56, 63, 64, 66, 67, 68, 69, 78, 79, 88, 94, 98,
     99, 101, 104, 105, 109, 110, 111, 112, 113, 114, 115, 116, 118, 119, 120, 121, 124, 125, 128, 129, 132, 134, 136,
     143, 144},

    {5, 11, 12, 14, 15, 16, 17, 24, 26, 27, 30, 32, 33, 38, 43, 45, 46, 47, 48, 49, 51, 52, 58, 59, 62, 66, 67, 68, 69,
     78, 79, 88, 89, 92, 93, 95, 96, 101, 104, 105, 107, 110, 111, 112, 113, 114, 115, 122, 126, 135, 136, 143},

    {1, 3, 5, 7, 9, 21, 22, 24, 25, 27, 30, 32, 33, 35, 36, 37, 39, 40, 44, 50, 52, 54, 58, 59, 65, 70, 71, 72, 73, 80,
     81, 82, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 95, 96, 101, 109, 133, 135, 136, 142, 143, 144},

    {2, 8, 9, 13, 14, 16, 17, 19, 24, 28, 30, 32, 34, 44, 45, 47, 48, 49, 51, 55, 58, 59, 60, 62, 76, 77, 84, 85, 86,
     87, 88, 92, 93, 94, 96, 101, 106, 108, 110, 111, 112, 113, 114, 115, 119, 127, 130, 133, 134, 143, 144},

    {14, 15, 16, 17, 24, 26, 28, 32, 33, 35, 36, 37, 43, 45, 46, 49, 51, 52, 54, 59, 62, 63, 64, 65, 66, 67, 68, 69, 84,
     85, 88, 89, 90, 91, 92, 93, 95, 96, 101, 110, 111, 122, 130, 135, 136, 139, 143},

    {11, 12, 14, 16, 17, 24, 26, 27, 29, 32, 33, 38, 43, 46, 47, 48, 49, 51, 53, 58, 59, 62, 65, 72, 73, 78, 79, 88, 89,
     92, 93, 95, 96, 101, 104, 105, 107, 112, 113, 114, 115, 126, 131, 135, 136, 143, 144},

    {14, 16, 17, 24, 26, 28, 32, 33, 35, 36, 37, 45, 49, 51, 52, 54, 59, 62, 63, 64, 65, 66, 67, 68, 69, 84, 85, 88, 90,
     91, 92, 93, 95, 96, 101, 110, 111, 135, 136, 140, 143},

]

target_paths = [set(path) for path in targetPaths]


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0


def compute_path_similarity_matrix(paths):
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix


def group_paths_by_similarity(paths):
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.mean(avg_sim_scores)

    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]
    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)

    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group


# === random grouping ===
USE_KEYBOARD_INPUT_GROUP_SIZE = False
RANDOM_GROUP_SEED = None


def group_paths_randomly(paths, use_keyboard_input=False, seed=None):
    n_paths = len(paths)
    original_group1, original_group2 = group_paths_by_similarity(paths)
    default_group1_size = len(original_group1)

    group1_size = default_group1_size

    if use_keyboard_input:
        while True:
            user_input = input(
                f"Random group1Number of Paths,  1~{n_paths - 1}; "
                f" {default_group1_size}: "
            ).strip()
            if user_input == "":
                group1_size = default_group1_size
                break
            try:
                group1_size = int(user_input)
                if 1 <= group1_size <= n_paths - 1:
                    break
                print(f": Random group1 1~{n_paths - 1} .")
            except ValueError:
                print(": , .")

    if not (1 <= group1_size <= n_paths - 1):
        raise ValueError(f"Random group1 1~{n_paths - 1} ,  {group1_size}")

    all_indices = list(range(n_paths))
    rng = random.Random(seed) if seed is not None else random
    rng.shuffle(all_indices)

    random_group1 = sorted(all_indices[:group1_size])
    random_group2 = sorted(all_indices[group1_size:])

    return random_group1, random_group2, default_group1_size, len(original_group2)


def compute_robustness(state, path):
    dx, dy, dz = state
    base = execute_Tr(dx, dy, dz)
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dw in [-1, 0, 1]:
        for dt in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dw == dt == dz == 0:
                    continue
                neighbor = np.clip(np.array(state) + np.array([dw, dt, dz]),
                                   STATE_MIN, STATE_MAX)
                neighbor = tuple(neighbor)
                ndx, ndy, ndz = neighbor
                n_trig = execute_Tr(ndx, ndy, ndz)
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def compute_q_value_score(state, similar_model):
    if similar_model is None:
        return 0.0
    try:
        normalized_state = normalizer.normalize(state)
        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = similar_model(state_tensor)
            max_q_value = torch.max(q_values).item()
            normalized_q = max_q_value / 20.0
            normalized_q = max(0.0, min(1.0, normalized_q))
            return 1.0 - normalized_q
    except:
        return 0.0


def generate_samples_for_similar_paths(similar_group, num_candidates=2000, top_k=200, run_id=1):
    SIMILAR_WEIGHTS = [0.55, 0.39, 0.06]

    def save_samples(path_id, samples, base_dir, group_type="similar"):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_{group_type}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"{group_type.title()} Group Path {path_id} - Run {run_id}\n")
            f.write("dx dy dz\tScore\tSimilarity\tLengthDiff\tRobustness\n")
            for s in samples:
                dx, dy, dz = s['state']
                f.write(
                    f"{dx} {dy} {dz}\t{s['score']:.4f}\t{s['similarity']:.4f}\t{s['length_diff']:.4f}\t{s['robustness']:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_grouped"

    for path_idx in similar_group:
        path = target_paths[path_idx]
        path_id = path_idx + 1
        candidate_samples = []
        attempts = 0

        while len(candidate_samples) < num_candidates and attempts < num_candidates * 10:
            attempts += 1
            state = random_state()
            dx, dy, dz = state
            triggered = execute_Tr(dx, dy, dz)
            if not triggered:
                continue
            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            candidate_samples.append({
                'state': state,
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            })

        if candidate_samples:
            for sample in candidate_samples:
                score = (SIMILAR_WEIGHTS[0] * sample['similarity'] +
                         SIMILAR_WEIGHTS[1] * sample['length_diff'] +
                         SIMILAR_WEIGHTS[2] * sample['robustness'])
                sample['score'] = score
            candidate_samples.sort(key=lambda x: x['score'], reverse=True)
            selected_samples = candidate_samples[:top_k]
            save_samples(path_id=path_id, samples=selected_samples, base_dir=base_dir, group_type="similar")


def generate_samples_for_isolated_paths(isolated_group, similar_model, num_candidates=2000, top_k=200, run_id=1):
    ISOLATED_WEIGHTS = [0.18, 0.21, 0.32, 0.29]

    def save_samples(path_id, samples, base_dir, group_type="isolated"):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_{group_type}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"{group_type.title()} Group Path {path_id} - Run {run_id}\n")
            f.write("dx dy dz\tScore\tSimilarity\tLengthDiff\tRobustness\tQValueScore\n")
            for s in samples:
                dx, dy, dz = s['state']
                f.write(
                    f"{dx} {dy} {dz}\t{s['score']:.4f}\t{s['similarity']:.4f}\t{s['length_diff']:.4f}\t{s['robustness']:.4f}\t{s['q_value_score']:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_grouped"

    for path_idx in isolated_group:
        path = target_paths[path_idx]
        path_id = path_idx + 1
        candidate_samples = []
        attempts = 0

        while len(candidate_samples) < num_candidates and attempts < num_candidates * 10:
            attempts += 1
            state = random_state()
            dx, dy, dz = state
            triggered = execute_Tr(dx, dy, dz)
            if not triggered:
                continue
            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            q_score = compute_q_value_score(state, similar_model)
            candidate_samples.append({
                'state': state,
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'q_value_score': q_score,
                'triggered': triggered
            })

        if candidate_samples:
            for sample in candidate_samples:
                score = (ISOLATED_WEIGHTS[0] * sample['similarity'] +
                         ISOLATED_WEIGHTS[1] * sample['length_diff'] +
                         ISOLATED_WEIGHTS[2] * sample['robustness'] +
                         ISOLATED_WEIGHTS[3] * sample['q_value_score'])
                sample['score'] = score
            candidate_samples.sort(key=lambda x: x['score'], reverse=True)
            selected_samples = candidate_samples[:top_k]
            save_samples(path_id=path_id, samples=selected_samples, base_dir=base_dir, group_type="isolated")


class GroupExperienceReplay:
    def __init__(self, capacity=20000):
        self.capacity = capacity
        self.buffer = deque(maxlen=self.capacity)
        self.priorities = deque(maxlen=self.capacity)
        self.sampled_indices = set()

    def append(self, experience):
        self.buffer.append(experience)
        self.priorities.append(experience[-1])

    def sample(self, batch_size, alpha=0.6):
        priorities = np.array(self.priorities) ** alpha
        probabilities = priorities / np.sum(priorities)
        batch_indices = np.random.choice(len(self.buffer), batch_size, p=probabilities)
        batch = [self.buffer[idx] for idx in batch_indices]
        return batch, batch_indices, probabilities[batch_indices]

    def update_priorities(self, batch_indices, td_errors):
        for idx, td_error in zip(batch_indices, td_errors):
            if idx < len(self.priorities):
                self.priorities[idx] = max(td_error, 1e-6)

    def __len__(self):
        return len(self.buffer)

    def get_high_reward_samples(self, target_path, num_samples=20):
        if len(self.buffer) == 0:
            return []

        samples_with_recalculated_scores = []
        for idx, experience in enumerate(self.buffer):
            if idx in self.sampled_indices:
                continue
            normalized_state_tensor = experience[0]
            normalized_state = normalized_state_tensor.cpu().numpy().flatten()
            state_tuple = tuple(normalizer.denormalize(normalized_state))
            dx, dy, dz = state_tuple
            triggered = execute_Tr(dx, dy, dz)
            new_reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_recalculated_scores.append((idx, state_tuple, new_reward, sim, triggered))

        samples_with_recalculated_scores.sort(key=lambda x: x[2], reverse=True)
        selected = samples_with_recalculated_scores[:num_samples]
        for item in selected:
            self.sampled_indices.add(item[0])
        return [(s[1], s[2], s[3], s[4]) for s in selected]

    def reset_sampled_indices(self):
        self.sampled_indices.clear()


def load_path_data(file_path):
    path_data = []
    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines[2:]:
            parts = line.strip().split("\t")
            state = tuple(map(int, parts[0].split()))
            path_data.append(state)
    return path_data


class DQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, 128)
        self.fc2 = nn.Linear(128, 64)
        self.fc3 = nn.Linear(64, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)


class DQNAgentWithPER:
    def __init__(self, state_dim, action_dim, replay_buffer, gamma=0.99, epsilon=1.0, epsilon_decay=0.995,
                 epsilon_min=0.1, learning_rate=0.001, alpha=0.6, beta=0.4):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer
        self.alpha = alpha
        self.beta = beta

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        delta_values = [1, -1]
        dim = action_idx // 2
        delta_idx = action_idx % 2
        delta = delta_values[delta_idx]
        if dim == 0:
            return (delta, 0, 0)
        elif dim == 1:
            return (0, delta, 0)
        elif dim == 2:
            return (0, 0, delta)

    def act(self, normalized_state):
        if random.random() < self.epsilon:
            return random.randrange(self.action_dim)
        state = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state)
        return torch.argmax(q_values, dim=1).item()

    def store_transition(self, normalized_state, action, reward, normalized_next_state, done):
        state = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        next_state = torch.tensor(normalized_next_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state)
            next_q_values = self.target_model(next_state)
            max_next_q_values = next_q_values.max(1)[0]
            target_q_values = reward + (self.gamma * max_next_q_values * (1 - done))
            td_error = torch.abs(q_values[0][action] - target_q_values).item()
        self.replay_buffer.append((state, action, reward, next_state, done, td_error))
        return td_error

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return
        batch, batch_indices, probabilities = self.replay_buffer.sample(batch_size, alpha=self.alpha)
        states, actions, rewards, next_states, dones, _ = zip(*batch)
        weights = (len(self.replay_buffer) * probabilities) ** (-self.beta)
        weights = weights / weights.max()
        weights = torch.tensor(weights, dtype=torch.float32).to(device)

        states = torch.tensor(np.array([s.cpu().numpy().flatten() for s in states]), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array([ns.cpu().numpy().flatten() for ns in next_states]),
                                   dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))
        td_errors = current_q_values - target_q_values
        weighted_loss = (td_errors.pow(2) * weights).mean()

        self.optimizer.zero_grad()
        weighted_loss.backward()
        self.optimizer.step()

        new_priorities = torch.abs(td_errors).detach().cpu().numpy()
        self.replay_buffer.update_priorities(batch_indices, new_priorities)

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())


def train_group(group_paths, path_documents, replay_buffer, batch_size=32, group_name="", pretrained_model=None,
                sample_group_type=None):
    if sample_group_type is None:
        sample_group_type = 'similar' if group_name == '' else 'isolated'
    state_dim = 3
    action_dim = 6

    agent = DQNAgentWithPER(state_dim, action_dim, replay_buffer)

    if pretrained_model is not None:
        print(f"  {group_name}: ()...")
        agent.model.load_state_dict(pretrained_model.state_dict())
        agent.target_model.load_state_dict(pretrained_model.state_dict())
        print(f"  {group_name}: completed")

    path_rewards = {}

    print(f"Start training{group_name}, Included Paths: {[idx + 1 for idx in group_paths]}")
    start_time = time.time()

    BATCH_SIZE = 50
    N_SAMPLES = 200
    N_STEPS = 3
    N_ROUNDS = 5
    N_BATCHES = 4

    replay_count = 0

    for path_idx in group_paths:
        file_path = os.path.join(path_documents,
                                 f"path{path_idx + 1}_{sample_group_type}.txt")
        if not os.path.exists(file_path):
            print(f"    : Path {path_idx + 1}, ")
            continue

        path_data = load_path_data(file_path)
        target_path = target_paths[path_idx]

        if path_idx not in path_rewards:
            path_rewards[path_idx] = 0

        print(f"\n  Start training path  {path_idx + 1},  {N_ROUNDS} ")

        for round_idx in range(N_ROUNDS):
            print(f"    Path  {path_idx + 1} - Run  {round_idx + 1}/{N_ROUNDS} ")

            for batch_idx in range(N_BATCHES):
                batch_start = batch_idx * BATCH_SIZE
                batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)

                if batch_start >= len(path_data):
                    print(f"       {batch_idx + 1}: , ")
                    break

                print(f"       {batch_idx + 1}/{N_BATCHES} ( {batch_start}-{batch_end})")

                for sample_idx in range(batch_start, batch_end):
                    if sample_idx >= len(path_data):
                        break

                    state = path_data[sample_idx]
                    prev_state = None
                    prev_triggered = None

                    for step in range(N_STEPS):
                        normalized_state = normalizer.normalize(state)

                        legal_actions = []
                        for a in range(agent.action_dim):
                            dw, dt, dz = agent.decode_action(a)
                            cand_next = tuple(np.clip(np.array(state) + np.array([dw, dt, dz]),
                                                      STATE_MIN, STATE_MAX))
                            legal_actions.append(a)

                        if not legal_actions:
                            break

                        if random.random() < agent.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        dw, dt, dz = agent.decode_action(action)
                        next_state = tuple(np.clip(np.array(state) + np.array([dw, dt, dz]),
                                                   STATE_MIN, STATE_MAX))

                        normalized_next_state = normalizer.normalize(next_state)

                        dx, dy, dz = next_state
                        triggered = execute_Tr(dx, dy, dz)
                        reward = compute_reward(next_state, target_path, triggered,
                                                prev_triggered, prev_state)
                        done = (step == N_STEPS - 1)

                        agent.store_transition(normalized_state, action, reward, normalized_next_state, done)

                        prev_state = state
                        prev_triggered = triggered
                        state = next_state
                        path_rewards[path_idx] += reward

                if len(agent.replay_buffer) >= batch_size:
                    agent.train(batch_size)
                    replay_count += 1
                    if replay_count % 2 == 0:
                        agent.update_target_model()

            print(f"      Path  {path_idx + 1} - Run  {round_idx + 1} completed")

        print(f"  Path  {path_idx + 1}  {N_ROUNDS} All completed ")

    training_time = time.time() - start_time
    print(f"\n{group_name}completed!")
    print(f"  : Path completed{N_ROUNDS}()")
    print(f"  : {replay_count}")
    print(f"  : {training_time:.2f} seconds")
    print(f"  : {len(replay_buffer)}")

    return agent, path_rewards, training_time


def generate_and_train_grouped_paths_staged(path_documents, random_group1, random_group2, batch_size=32, run_id=1):
    print(f"\n===  {run_id}/20 (3 minutes, random grouping+model reuse) ===")
    random_group1_paths = [idx + 1 for idx in random_group1]
    random_group2_paths = [idx + 1 for idx in random_group2]

    print(f"Random group1Path (pretrained group): {random_group1_paths}")
    print(f"Random group2Path (model-reuse group): {random_group2_paths}")

    total_start_time = time.time()

    print(f"\n[1] Random group1...")
    generate_samples_for_similar_paths(random_group1, num_candidates=2000, top_k=200, run_id=run_id)

    print(f"\n[2] Random group1(, {5})...")
    group1_replay_buffer = GroupExperienceReplay(capacity=20000)
    group1_agent, group1_path_rewards, group1_training_time = train_group(
        random_group1, path_documents, group1_replay_buffer, batch_size,
        group_name="Random group1(pretrained group)", pretrained_model=None, sample_group_type="similar"
    )

    print(f"\n[3] Random group1Random group2...")
    generate_samples_for_isolated_paths(random_group2, group1_agent.model,
                                        num_candidates=2000, top_k=200, run_id=run_id)

    print(f"\n[4] Random group2(Random group1, {5})...")
    group2_replay_buffer = GroupExperienceReplay(capacity=20000)
    group2_agent, group2_path_rewards, group2_training_time = train_group(
        random_group2, path_documents, group2_replay_buffer, batch_size,
        group_name="Random group2(model-reuse group)", pretrained_model=group1_agent.model, sample_group_type="isolated"
    )

    total_path_rewards = {**group1_path_rewards, **group2_path_rewards}
    total_cumulative_reward = sum(total_path_rewards.values())
    total_training_time = time.time() - total_start_time

    print(f"\n===  {run_id}/20 completed, : {total_training_time:.2f} seconds ===")
    print(f"Random group1: {group1_training_time:.2f} seconds")
    print(f"Random group2: {group2_training_time:.2f} seconds")
    print(f" - Random group1: {len(group1_replay_buffer)}, Random group2: {len(group2_replay_buffer)}")

    return group1_agent, group2_agent, group1_replay_buffer, group2_replay_buffer, \
        total_cumulative_reward, total_path_rewards, total_training_time


def create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir):
    os.makedirs(output_dir, exist_ok=True)

    similar_group_paths = [idx + 1 for idx in similar_group]
    isolated_group_paths = [idx + 1 for idx in isolated_group]

    wb = Workbook()

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    header_color = "4472C4"
    similar_group_color = "E2EFDA"
    isolated_group_color = "FCE4D6"
    stats_color = "FFF2CC"

    ws_paths = wb.active
    ws_paths.title = "Path "

    path_headers = ['Path ID', ''] + [f'Run {i}' for i in range(1, 21)] + ['Average Similarity', 'Maximum Similarity',
                                                                                    'Minimum Similarity', 'Standard deviation']
    for col, header in enumerate(path_headers, 1):
        cell = ws_paths.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_paths.row_dimensions[1].height = 30

    for path_id in range(1, len(target_paths) + 1):
        row = path_id + 1

        if path_id in similar_group_paths:
            group_type = "Random group1(pretrained group)"
            row_color = similar_group_color
        elif path_id in isolated_group_paths:
            group_type = "Random group2(model-reuse group)"
            row_color = isolated_group_color
        else:
            group_type = "Ungrouped"
            row_color = "FFFFFF"

        cell = ws_paths.cell(row=row, column=1, value=f"Path {path_id}")
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        cell.border = thin_border

        cell = ws_paths.cell(row=row, column=2, value=group_type)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        cell.border = thin_border

        path_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            sim = run_data['path_similarities'].get(path_id, {}).get('avg_similarity', 0.0)
            path_similarities.append(sim)

            cell = ws_paths.cell(row=row, column=3 + run_idx, value=round(sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        stats_values = [
            round(np.mean(path_similarities), 4),
            round(np.max(path_similarities), 4),
            round(np.min(path_similarities), 4),
            round(np.std(path_similarities), 4)
        ]

        for i, value in enumerate(stats_values):
            cell = ws_paths.cell(row=row, column=23 + i, value=value)
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
            cell.font = Font(bold=True, size=10)
            cell.border = thin_border

    ws_paths.column_dimensions['A'].width = 13
    ws_paths.column_dimensions['B'].width = 16
    for col in range(3, 23):
        ws_paths.column_dimensions[get_column_letter(col)].width = 10
    for col in range(23, 27):
        ws_paths.column_dimensions[get_column_letter(col)].width = 13

    ws_groups = wb.create_sheet("")

    group_headers = ['Group Name', 'Included Paths'] + [f'Run {i}' for i in range(1, 21)] + ['Average Similarity', 'Standard deviation']
    for col, header in enumerate(group_headers, 1):
        cell = ws_groups.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_groups.row_dimensions[1].height = 30

    row = 2

    cell = ws_groups.cell(row=row, column=1, value="Random group1(pretrained group)")
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    cell.border = thin_border

    cell = ws_groups.cell(row=row, column=2, value=','.join(map(str, similar_group_paths)))
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    cell.border = thin_border

    group_similarities = []
    for run_idx, run_data in enumerate(all_runs_data):
        group_sim = np.mean(
            [run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0) for p in similar_group_paths])
        group_similarities.append(group_sim)

        cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(group_sim, 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    cell = ws_groups.cell(row=row, column=23, value=round(np.mean(group_similarities), 4))
    cell.number_format = '0.0000'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border

    cell = ws_groups.cell(row=row, column=24, value=round(np.std(group_similarities), 4))
    cell.number_format = '0.0000'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border

    row += 1

    if isolated_group_paths:
        cell = ws_groups.cell(row=row, column=1, value="Random group2(model-reuse group)")
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        cell.border = thin_border

        cell = ws_groups.cell(row=row, column=2, value=','.join(map(str, isolated_group_paths)))
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        cell.border = thin_border

        isolated_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            iso_sim = np.mean(
                [run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0) for p in isolated_group_paths])
            isolated_similarities.append(iso_sim)

            cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(iso_sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        cell = ws_groups.cell(row=row, column=23, value=round(np.mean(isolated_similarities), 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border

        cell = ws_groups.cell(row=row, column=24, value=round(np.std(isolated_similarities), 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border

    ws_groups.column_dimensions['A'].width = 16
    ws_groups.column_dimensions['B'].width = 22
    for col in range(3, 23):
        ws_groups.column_dimensions[get_column_letter(col)].width = 10
    ws_groups.column_dimensions[get_column_letter(23)].width = 14
    ws_groups.column_dimensions[get_column_letter(24)].width = 12

    ws_samples = wb.create_sheet("Detailed Sample Data")

    sample_headers = ['Run', 'Path ID', 'Sample ID', 'Dx', 'Dy', 'Dz', 'Similarity', 'Triggered Rule Set']
    for col, header in enumerate(sample_headers, 1):
        cell = ws_samples.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_samples.row_dimensions[1].height = 30

    sample_row = 2
    for run_idx, run_data in enumerate(all_runs_data, 1):
        for path_id in range(1, len(target_paths) + 1):
            samples = run_data['path_samples'].get(path_id, [])

            if path_id in similar_group_paths:
                path_color = similar_group_color
            elif path_id in isolated_group_paths:
                path_color = isolated_group_color
            else:
                path_color = "FFFFFF"

            for sample_idx, (state_tuple, reward, sim, triggered) in enumerate(samples, 1):
                dx, dy, dz = state_tuple
                triggered_str = ','.join(map(str, sorted(triggered)))

                cell = ws_samples.cell(row=sample_row, column=1, value=f"Run {run_idx}")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=path_color, end_color=path_color, fill_type="solid")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=2, value=f"Path {path_id}")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=path_color, end_color=path_color, fill_type="solid")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=3, value=sample_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                for col_offset, value in enumerate([dx, dy, dz]):
                    cell = ws_samples.cell(row=sample_row, column=4 + col_offset, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=7, value=round(sim, 4))
                cell.number_format = '0.0000'
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=8, value=f"{{{triggered_str}}}")
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border

                sample_row += 1

    sample_widths = [13, 13, 11, 10, 12, 8, 12, 45]
    for i, width in enumerate(sample_widths, 1):
        ws_samples.column_dimensions[get_column_letter(i)].width = width

    output_path = os.path.join(output_dir, "20 run_random grouping_model reuse_3 minutes.xlsx")
    wb.save(output_path)
    print(f"\n Consolidated Excel report generated: {output_path}")


def run_20_times_training():
    model_path_base = r"D:\Experiment\CNN\DQNNEW\saved_models_random_reuse_3min_version"
    path_documents = r"D:\Experiment\CNN\DQNNEW\path_samples_grouped"
    output_dir = r"D:\Experiment\CNN\ComparisonExperiment2\excel_reports_random_reuse_3min_version"

    os.makedirs(model_path_base, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    similar_group, isolated_group, default_group1_size, default_group2_size = group_paths_randomly(
        target_paths,
        use_keyboard_input=USE_KEYBOARD_INPUT_GROUP_SIZE,
        seed=RANDOM_GROUP_SEED
    )

    similar_group_display = [idx + 1 for idx in similar_group]
    isolated_group_display = [idx + 1 for idx in isolated_group]

    print("=" * 60)
    print("20 - random grouping + model reuse - 3 minutes")
    print("=" * 60)
    print("Training-scale configuration:")
    print("   Per path: 5")
    print("   Per round: 4")
    print("   Per batch: 50")
    print("   Per sample: 3")
    print("   Sample generation: 2000candidates -> 200final samples")
    print("   : save model parameters only(optimized version)")
    print("   : ")
    print(f"   Default group size: Random group1={default_group1_size}Path , Random group2={default_group2_size}Path ")
    print(f"   Keyboard-input group size: {'' if USE_KEYBOARD_INPUT_GROUP_SIZE else ''}")
    print(f"   Random seed: {RANDOM_GROUP_SEED if RANDOM_GROUP_SEED is not None else 'None,  run'}")
    print("=" * 60)
    print(f"\nAutomatic grouping results:")
    print(f"Similar path group: {similar_group_display}")
    print(f"Isolated path group: {isolated_group_display}")
    print("\n" + "=" * 60)

    all_runs_data = []
    total_start_time = time.time()

    for run_id in range(1, 21):
        print(f"\n{'=' * 60}")
        print(f"Start run  {run_id}/20  run")
        print(f"{'=' * 60}")

        group1_agent, group2_agent, group1_buffer, group2_buffer, total_cumulative_reward, path_rewards, training_time = \
            generate_and_train_grouped_paths_staged(path_documents, similar_group, isolated_group, batch_size=32,
                                                    run_id=run_id)

        group1_model_path = os.path.join(model_path_base, f"random_group1_model_run_{run_id}.pth")
        group2_model_path = os.path.join(model_path_base, f"random_group2_model_run_{run_id}.pth")

        torch.save(group1_agent.model.state_dict(), group1_model_path)
        torch.save(group2_agent.model.state_dict(), group2_model_path)

        print(f"[Run {run_id}] Model saved(optimized version - )")

        group1_buffer.reset_sampled_indices()
        group2_buffer.reset_sampled_indices()

        run_data = {
            'run_id': run_id,
            'training_time': training_time,
            'total_reward': total_cumulative_reward,
            'path_rewards': path_rewards,
            'path_similarities': {},
            'path_samples': {}
        }

        all_similarities = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            path_id = path_idx + 1

            if path_id in similar_group_display:
                buffer = group1_buffer
            elif path_id in isolated_group_display:
                buffer = group2_buffer
            else:
                continue

            high_reward_samples = buffer.get_high_reward_samples(target_path, num_samples=20)

            if high_reward_samples:
                similarities = [sim for _, _, sim, _ in high_reward_samples]
                run_data['path_similarities'][path_idx + 1] = {
                    'avg_similarity': np.mean(similarities),
                    'max_similarity': np.max(similarities),
                    'min_similarity': np.min(similarities),
                    'sample_count': len(similarities)
                }
                run_data['path_samples'][path_idx + 1] = high_reward_samples
                all_similarities.extend(similarities)
            else:
                run_data['path_similarities'][path_idx + 1] = {
                    'avg_similarity': 0.0,
                    'max_similarity': 0.0,
                    'min_similarity': 0.0,
                    'sample_count': 0
                }
                run_data['path_samples'][path_idx + 1] = []

        if all_similarities:
            run_data['overall_avg_similarity'] = np.mean(all_similarities)
            run_data['max_similarity'] = np.max(all_similarities)
            run_data['min_similarity'] = np.min(all_similarities)
        else:
            run_data['overall_avg_similarity'] = 0.0
            run_data['max_similarity'] = 0.0
            run_data['min_similarity'] = 0.0

        all_runs_data.append(run_data)

        print(f"[Run {run_id}] completed! Overall Average Similarity: {run_data['overall_avg_similarity']:.4f}")
        print(f"{'=' * 60}\n")

    total_time = time.time() - total_start_time

    print("\nGenerating consolidated Excel report...")
    create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir)

    print("\n" + "=" * 60)
    print("20All completed! - random grouping + model reuse - 3 minutes")
    print("=" * 60)
    print(f":")
    print(f"  Per path: 5 x 4 x 50 x 3 = 3000/Path ")
    print(f"  Sample generation: 2000candidates -> 200final samples")
    print(f"  : save model parameters only(optimized version)")
    print(f"  Total elapsed time: {total_time:.2f} seconds ({total_time / 60:.2f} minutes)")
    print(f"  Average elapsed time per run: {total_time / 20:.2f} seconds")
    print(f"\nAverage similarity statistics:")
    avg_similarities = [r['overall_avg_similarity'] for r in all_runs_data]
    print(f"  Overall average: {np.mean(avg_similarities):.4f}")
    print(f"  Maximum: {np.max(avg_similarities):.4f}")
    print(f"  Minimum: {np.min(avg_similarities):.4f}")
    print(f"  Standard deviation: {np.std(avg_similarities):.4f}")
    print(f"\nAll results have been saved to: {output_dir}")
    print("=" * 60)


if __name__ == "__main__":
    run_20_times_training()