import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

def execute_validation_rules(x, y, z):
    """第8类: 速度扭矩电流混合控制 (速度x, 扭矩y, 电流z) - 144个有效变异分支"""
    triggered = set()  # 记录触发的条件编号（1-144）


    # -------------------------- 1-32：速度扭矩电流基础检查 --------------------------
    # 编号1 (原编号7)
    if [(x > 1400 and x < 1600 and y > 160 and y < 190)] != [(x > 1400 and x * 8 < 1600 and y > 160 and y < 190)]:
        triggered.add(1)
    # 编号2 (原编号8)
    if [(x > 1400 and x < 1600 and y > 160 and y < 190)] != [(x > 1400 and x < 1600 and y * 10 > 160 and y < 190)]:
        triggered.add(2)
    # 编号3 (原编号9)
    if [(x > 1350 and x < 1650 and z > 11 and z < 14)] != [(x > 1350 and x * 9 < 1650 and z > 11 and z < 14)]:
        triggered.add(3)
    # 编号4 (原编号10)
    if [(x > 1350 and x < 1650 and z > 11 and z < 14)] != [(x > 1350 and x < 1650 and z * 20 > 11 and z < 14)]:
        triggered.add(4)
    # 编号5 (原编号11)
    if [(y > 155 and y < 195 and z > 11 and z < 14)] != [(y > 155 and y * 7 < 195 and z > 11 and z < 14)]:
        triggered.add(5)
    # 编号6 (原编号12)
    if [(y > 155 and y < 195 and z > 11 and z < 14)] != [(y > 155 and y < 195 and z * 22 > 11 and z < 14)]:
        triggered.add(6)
    # 编号7 (原编号13)
    if [(x > 1480 and x < 1520)] != [(x > 1480 and x * 8 < 1520)]:
        triggered.add(7)
    # 编号8 (原编号14)
    if [(x > 1480 and x < 1520)] != [(x * 12 > 1480 and x < 1520)]:
        triggered.add(8)
    # 编号9 (原编号15)
    if [(y > 173 and y < 177)] != [(y * 12 > 173 and y < 177)]:
        triggered.add(9)
    # 编号10 (原编号16)
    if [(y > 173 and y < 177)] != [(y > 173 and y * 22 < 177)]:
        triggered.add(10)
    # 编号11 (原编号17)
    if [(z > 12.2 and z < 12.8)] != [(z * 78 > 12.2 and z < 12.8)]:
        triggered.add(11)
    # 编号12 (原编号18)
    if [(z > 12.2 and z < 12.8)] != [(66 > 12.2 and z < 12.8)]:
        triggered.add(12)
    # 编号13 (原编号19)
    if [(abs(x - 1500) < 40)] != [(abs(x - 1500) < 140)]:
        triggered.add(13)
    # 编号14 (原编号20)
    if [(abs(x - 1500) < 40)] != [(abs(x - 1500) < 990)]:
        triggered.add(14)
    # 编号15 (原编号21)
    if [(abs(y - 175) < 6)] != [(abs(y - 175) < 16)]:
        triggered.add(15)
    # 编号16 (原编号22)
    if [(abs(y - 175) < 6)] != [(abs(y - 175) < 226)]:
        triggered.add(16)
    # 编号17 (原编号23)
    if [(abs(z - 12.5) < 0.4)] != [(abs(z - 12.5) < 4)]:
        triggered.add(17)
    # 编号18 (原编号24)
    if [(abs(z - 12.5) < 0.4)] != [(abs(z - 12.5) < 1)]:
        triggered.add(18)
    # 编号19 (原编号25)
    if [(x > 1350 and y > 155 and z > 11)] != [(x > 1350 and y * 78 > 155 and z > 11)]:
        triggered.add(19)
    # 编号20 (原编号26)
    if [(x > 1350 and y > 155 and z > 11)] != [(x * 67 > 1350 and y > 155 and z > 11)]:
        triggered.add(20)
    # 编号21 (原编号27)
    if [(x < 1650 and y < 195 and z < 14)] != [(x < 1650 and y * 88 < 195 and z < 14)]:
        triggered.add(21)
    # 编号22 (原编号28)
    if [(x < 1650 and y < 195 and z < 14)] != [(x < 1650 and y < 195 and z * 8 < 14)]:
        triggered.add(22)
    # 编号23 (原编号29)
    if [((x / 10 + y + z * 10) > 350 and (x / 10 + y + z * 10) < 410)] != [
        ((x / 10 + y + z * 10) > 3150 and (x / 10 + y + z * 10) < 410)]:
        triggered.add(23)
    # 编号24 (原编号30)
    if [((x / 10 + y + z * 10) > 350 and (x / 10 + y + z * 10) < 410)] != [
        ((x / 10 + y + z * 10) > 350 and (x / 10 + y + z * 10) < 1410)]:
        triggered.add(24)
    # 编号25 (原编号31)
    if [(x * y / 1000 > 230 and x * y / 1000 < 310)] != [(x * y / 1000 > 1230 and x * y / 1000 < 310)]:
        triggered.add(25)
    # 编号26 (原编号32)
    if [(x * y / 1000 > 230 and x * y / 1000 < 310)] != [(x * y / 1000 > 230 and x * y / 1000 < 3110)]:
        triggered.add(26)
    # 编号27 (原编号33)
    if [(x * z / 100 > 160 and x * z / 100 < 220)] != [(x * z / 100 > 1610 and x * z / 100 < 220)]:
        triggered.add(27)
    # 编号28 (原编号34)
    if [(x * z / 100 > 160 and x * z / 100 < 220)] != [(x * z / 100 > 160 and x * z / 100 < 2120)]:
        triggered.add(28)
    # 编号29 (原编号35)
    if [(y * z > 2000 and y * z < 2600)] != [(y * z > 200 and y * z < 2600)]:
        triggered.add(29)
    # 编号30 (原编号36)
    if [(y * z > 2000 and y * z < 2600)] != [(y * z > 2000 and y * z < 260)]:
        triggered.add(30)
    # 编号31 (原编号37)
    if [((x / 10 + y + z * 10) / 3 > 116 and (x / 10 + y + z * 10) / 3 < 136)] != [
        ((x / 10 + y + z * 10) / 5 > 116 and (x / 10 + y + z * 10) / 3 < 136)]:
        triggered.add(31)
    # 编号32 (原编号38)
    if [((x / 10 + y + z * 10) / 3 > 116 and (x / 10 + y + z * 10) / 3 < 136)] != [
        ((x / 10 + y + z * 10) / 3 > 116 and (x / 10 + y + z * 10) / 11 < 136)]:
        triggered.add(32)

    # -------------------------- 33-58：驱动系统分析 --------------------------
    # 编号33 (原编号41)
    if [(x / (y * 10) > 0.8 and x / (y * 10) < 1.0)] != [(x / (y * 10) > 8 and x / (y * 10) < 1.0)]:
        triggered.add(33)
    # 编号34 (原编号42)
    if [(x / (y * 10) > 0.8 and x / (y * 10) < 1.0)] != [(x / (y * 10) > 0.8 and x / (y * 10) < 10)]:
        triggered.add(34)
    # 编号35 (原编号43)
    if [(x / z > 110 and x / z < 140)] != [(x / z > 1220 and x / z < 140)]:
        triggered.add(35)
    # 编号36 (原编号44)
    if [(x / z > 110 and x / z < 140)] != [(x / z > 110 and x / z < 14)]:
        triggered.add(36)
    # 编号37 (原编号45)
    if [(y / z > 12 and y / z < 16)] != [(y / z > 112 and y / z < 16)]:
        triggered.add(37)
    # 编号38 (原编号46)
    if [(y / z > 12 and y / z < 16)] != [(y / z > 12 and y / z < 1226)]:
        triggered.add(38)
    # 编号39 (原编号47)
    if [((x - 1200) / 10 + (y - 150) > 35 and (x - 1200) / 10 + (y - 150) < 55)] != [
        ((x - 1200) / 10 + (y - 150) > 315 and (x - 1200) / 10 + (y - 150) < 55)]:
        triggered.add(39)
    # 编号40 (原编号48)
    if [((x - 1200) / 10 + (y - 150) > 35 and (x - 1200) / 10 + (y - 150) < 55)] != [
        ((x - 1200) / 10 + (y - 220) > 35 and (x - 1200) / 10 + (y - 150) < 55)]:
        triggered.add(40)
    # 编号41 (原编号49)
    if [((x - 1200) / 10 + (z - 10) * 5 > 40 and (x - 1200) / 10 + (z - 10) * 5 < 60)] != [
        ((x - 1200) / 10 + (z - 10) * 15 > 40 and (x - 1200) / 10 + (z - 10) * 5 < 60)]:
        triggered.add(41)
    # 编号42 (原编号50)
    if [((x - 1200) / 10 + (z - 10) * 5 > 40 and (x - 1200) / 10 + (z - 10) * 5 < 60)] != [
        ((x - 1200) / 10 + (z - 10) * 115 > 40 and (x - 1200) / 10 + (z - 10) * 5 < 60)]:
        triggered.add(42)
    # 编号43 (原编号51)
    if [((y - 150) + (z - 10) * 2 > 30 and (y - 150) + (z - 10) * 2 < 50)] != [
        ((y - 150) + (z - 10) * 2 > 302 and (y - 150) + (z - 10) * 2 < 50)]:
        triggered.add(43)
    # 编号44 (原编号52)
    if [((y - 150) + (z - 10) * 2 > 30 and (y - 150) + (z - 10) * 2 < 50)] != [
        ((y - 150) + (z - 10) * 211 > 30 and (y - 150) + (z - 10) * 2 < 50)]:
        triggered.add(44)
    # 编号45 (原编号53)
    if [(abs((x - 1500) / 10 - (y - 175)) < 12)] != [(abs((x - 1500) / 10 - (y - 175)) < 112)]:
        triggered.add(45)
    # 编号46 (原编号54)
    if [(abs((x - 1500) / 10 - (y - 175)) < 12)] != [(abs((x - 1500) / 101 - (y - 175)) < 12)]:
        triggered.add(46)
    # 编号47 (原编号55)
    if [(abs((x - 1500) / 100 - (z - 12.5)) < 3)] != [(abs((x - 1500) / 100 - (z - 12.5)) < 43)]:
        triggered.add(47)
    # 编号48 (原编号56)
    if [(abs((x - 1500) / 100 - (z - 12.5)) < 3)] != [(abs((x - 1500) / 100 - (z - 12.5)) < 13)]:
        triggered.add(48)
    # 编号49 (原编号57)
    if [(abs((y - 175) - (z - 12.5) * 12) < 8)] != [(abs((y - 175) - (z - 12.5) * 12) < 118)]:
        triggered.add(49)
    # 编号50 (原编号58)
    if [(abs((y - 175) - (z - 12.5) * 12) < 8)] != [(abs((y - 175) - (z - 12.5) * 232) < 8)]:
        triggered.add(50)
    # 编号51 (原编号60)
    if [(x / (y * 10 + 500) > 0.7 and x / (y * 10 + 500) < 0.9)] != [
        (x / (y * 10 + 500) > 337 and x / (y * 10 + 500) < 0.9)]:
        triggered.add(51)
    # 编号52 (原编号61)
    if [(y / (z + 5) > 9 and y / (z + 5) < 13)] != [(y / (z + 5) > 119 and y / (z + 5) < 13)]:
        triggered.add(52)
    # 编号53 (原编号62)
    if [(y / (z + 5) > 9 and y / (z + 5) < 13)] != [(y / (z + 5) > 9 and y / (z + 5) < 123)]:
        triggered.add(53)
    # 编号54 (原编号63)
    if [(z / (x / 150) > 1.1 and z / (x / 150) < 1.5)] != [(z / (x / 150) > 11 and z / (x / 150) < 1.5)]:
        triggered.add(54)
    # 编号55 (原编号64)
    if [(z / (x / 150) > 1.1 and z / (x / 150) < 1.5)] != [(z / (x / 150) > 1.1 and z / (x / 150) < 15)]:
        triggered.add(55)
    # 编号56 (原编号65)
    if [((x * y * z / 10000) > 28 and (x * y * z / 10000) < 38)] != [
        ((x * y * z / 10000) > 28 and (x * y * z / 10000) < 238)]:
        triggered.add(56)
    # 编号57 (原编号67)
    if [(x / 20 + y * 0.6 + z * 8 > 260 and x / 20 + y * 0.6 + z * 8 < 300)] != [
        (x / 20 + y * 0.6 + z * 10 > 260 and x / 20 + y * 0.6 + z * 8 < 300)]:
        triggered.add(57)
    # 编号58 (原编号68)
    if [(x / 20 + y * 0.6 + z * 8 > 260 and x / 20 + y * 0.6 + z * 8 < 300)] != [
        (x / 20 + y * 0.6 + z * 8 > 260 and x / 20 + y * 6 + z * 8 < 300)]:
        triggered.add(58)
    # 编号59 (原编号69)
    if [((x / 1500) ** 0.4 * (y / 175) ** 0.4 * (z / 12.5) ** 0.2 > 0.93)] != [
        ((x / 1500) ** 0.4 * (y / 175) ** 0.4 * (z / 12.5) ** 0.2 > 93)]:
        triggered.add(59)
    # 编号60 (原编号70)
    if [((x / 1500) ** 0.4 * (y / 175) ** 0.4 * (z / 12.5) ** 0.2 > 0.93)] != [
        ((x / 1500) ** 0.4 * (y / 175) ** 4 * (z / 12.5) ** 0.2 > 0.93)]:
        triggered.add(60)
    # 编号61 (原编号71)
    if [((x - 1500) / 10 * (y - 175) > -200 and (x - 1500) / 10 * (y - 175) < 200)] != [
        ((x - 1500) / 110 * (y - 175) > -200 and (x - 1500) / 10 * (y - 175) < 200)]:
        triggered.add(61)
    # 编号62 (原编号72)
    if [((x - 1500) / 10 * (y - 175) > -200 and (x - 1500) / 10 * (y - 175) < 200)] != [
        ((x - 1500) / 10 * (y - 175) > -200 and (x - 1500) / 110 * (y - 175) < 200)]:
        triggered.add(62)
    # 编号63 (原编号73)
    if [((x - 1500) / 100 * (z - 12.5) > -20 and (x - 1500) / 100 * (z - 12.5) < 20)] != [
        ((x - 1500) / 100 * (z - 12.5) > -20 and (x - 1500) / 10 * (z - 12.5) < 20)]:
        triggered.add(63)
    # 编号64 (原编号75)
    if [((y - 175) * (z - 12.5) > -20 and (y - 175) * (z - 12.5) < 20)] != [
        ((y - 175) * (z - 12.5) > -20 and (y - 175) * (z - 12.5) < 210)]:
        triggered.add(64)
    # 编号65 (原编号76)
    if [((y - 175) * (z - 12.5) > -20 and (y - 175) * (z - 12.5) < 20)] != [((y - 175) * (z - 125) < 20)]:
        triggered.add(65)
    # 编号66 (原编号77)
    if [(abs(x / 1500 + y / 175 + z / 12.5 - 3) < 0.2)] != [(abs(x / 1500 + y / 175 + z / 12.5 - 3) < 112)]:
        triggered.add(66)
    # 编号67 (原编号78)
    if [(abs(x / 1500 + y / 175 + z / 12.5 - 3) < 0.2)] != [(abs(x / 1500 + y / 175 + z / 12.5 - 3) < 2222)]:
        triggered.add(67)
    # 编号68 (原编号79)
    if [(abs((x / 1500) * (y / 175) * (z / 12.5) - 1) < 0.15)] != [
        (abs((x / 1500) * (y / 175) * (z / 12.5) - 1) < 115)]:
        triggered.add(68)
    # 编号69 (原编号80)
    if [(abs((x / 1500) * (y / 175) * (z / 12.5) - 1) < 0.15)] != [
        (abs((x / 1500) * (y / 175) * (z / 12.5) - 1) < 2225)]:
        triggered.add(69)

    # -------------------------- 70-107：动力传动协调 --------------------------
    # 编号70 (原编号81)
    if [(x > 1480 and x < 1520 and y > 172 and y < 178 and z > 12.2 and z < 12.8)] != [
        (x > 1480 and x < 1520 and y > 172 and y < 178 and z > 12.2 and 8 < 12.8)]:
        triggered.add(70)
    # 编号71 (原编号82)
    if [(x > 1480 and x < 1520 and y > 172 and y < 178 and z > 12.2 and z < 12.8)] != [
        (x > 1480 and x < 1520 and y > 172 and y < 178 and z > 12.2 and 10 < 12.8)]:
        triggered.add(71)
    # 编号72 (原编号83)
    if [(abs(x / (y * 10) - 0.86) < 0.06 and abs(z - 12.5) < 0.3)] != [
        (abs(x / (y * 10) - 0.86) < 0.06 and abs(z - 12.5) < 3)]:
        triggered.add(72)
    # 编号73 (原编号84)
    if [(abs(x / (y * 10) - 0.86) < 0.06 and abs(z - 12.5) < 0.3)] != [
        (abs(x / (y * 10) - 0.86) < 0.06 and abs(z - 12.5) < 13)]:
        triggered.add(73)
    # 编号74 (原编号85)
    if [(x * y / 1000 > 258 and x * y / 1000 < 268 and z > 12.3 and z < 12.7)] != [
        (x * y / 1000 > 258 and x * y / 1000 < 268 and 8 * z > 12.3 and z < 12.7)]:
        triggered.add(74)
    # 编号75 (原编号86)
    if [(x * y / 1000 > 258 and x * y / 1000 < 268 and z > 12.3 and z < 12.7)] != [
        (x * y / 1000 > 258 and x * y / 1000 < 268 and z * 99 > 12.3 and z < 12.7)]:
        triggered.add(75)
    # 编号76 (原编号87)
    if [((x / 10 + y) / 2 > 137 and (x / 10 + y) / 2 < 143 and z > 12.2 and z < 12.8)] != [
        ((x / 10 + y) / 2 > 137 and (x / 10 + y) / 2 < 143 and z > 12.2 and 8 < 12.8)]:
        triggered.add(76)
    # 编号77 (原编号88)
    if [((x / 10 + y) / 2 > 137 and (x / 10 + y) / 2 < 143 and z > 12.2 and z < 12.8)] != [
        ((x / 10 + y) / 2 > 137 and (x / 10 + y) / 2 < 143 and z > 12.2 and 10 < 12.8)]:
        triggered.add(77)
    # 编号78 (原编号89)
    if [(abs(x / 10 - y) < 25 and z > 12.2 and z < 12.8)] != [(abs(x / 10 - y) < 25 and z * 12 > 12.2 and z < 12.8)]:
        triggered.add(78)
    # 编号79 (原编号90)
    if [(abs(x / 10 - y) < 25 and z > 12.2 and z < 12.8)] != [(abs(x / 10 - y) < 25 and z * 89 > 12.2 and z < 12.8)]:
        triggered.add(79)
    # 编号80 (原编号91)
    if [(math.sqrt((x / 10 - 150) ** 2 + (y - 175) ** 2) < 6 and abs(z - 12.5) < 0.3)] != [
        (math.sqrt((x / 10 - 150) ** 2 + (y - 175) ** 2) < 6 and abs(z - 12.5) < 3)]:
        triggered.add(80)
    # 编号81 (原编号92)
    if [(math.sqrt((x / 10 - 150) ** 2 + (y - 175) ** 2) < 6 and abs(z - 12.5) < 0.3)] != [
        (math.sqrt((x / 10 - 150) ** 2 + (y - 175) ** 2) < 6 and abs(z - 12.5) < 13)]:
        triggered.add(81)
    # 编号82 (原编号93)
    if [(x / 1500 > 0.98 and x / 1500 < 1.02 and y / 175 > 0.97 and y / 175 < 1.03)] != [
        (x / 1500 > 0.98 and x / 1500 < 1.02 and y / 175 > 97 and y / 175 < 1.03)]:
        triggered.add(82)
    # 编号83 (原编号94)
    if [(x / 1500 > 0.98 and x / 1500 < 1.02 and y / 175 > 0.97 and y / 175 < 1.03)] != [
        (x / 1500 > 0.98 and x / 1500 < 1.02 and y * 8 / 175 > 0.97 and y / 175 < 1.03)]:
        triggered.add(83)
    # 编号84 (原编号95)
    if [(z / 12.5 > 0.976 and z / 12.5 < 1.024)] != [(z / 12.5 > 0.976 and z / 125 < 1.024)]:
        triggered.add(84)
    # 编号85 (原编号96)
    if [(z / 12.5 > 0.976 and z / 12.5 < 1.024)] != [(z / 12.5 > 0.976 and z / 200 < 1.024)]:
        triggered.add(85)
    # 编号86 (原编号97)
    if [((x / 1500 + y / 175 + z / 12.5) / 3 > 0.98 and (x / 1500 + y / 175 + z / 12.5) / 3 < 1.02)] != [
        ((x / 1500 + y / 175 + z / 12.5) / 3 > 98 and (x / 1500 + y / 175 + z / 12.5) / 3 < 1.02)]:
        triggered.add(86)
    # 编号87 (原编号98)
    if [((x / 1500 + y / 175 + z / 12.5) / 3 > 0.98 and (x / 1500 + y / 175 + z / 12.5) / 3 < 1.02)] != [
        ((x / 1500 + y / 175 + z / 12.5) / 23 > 0.98 and (x / 1500 + y / 175 + z / 12.5) / 3 < 1.02)]:
        triggered.add(87)
    # 编号88 (原编号99)
    if [(max(abs(x / 1500 - 1), abs(y / 175 - 1), abs(z / 12.5 - 1)) < 0.025)] != [
        (max(abs(x / 1500 - 1), abs(y / 175 - 1), abs(z / 12.5 - 1)) < 0.55)]:
        triggered.add(88)
    # 编号89 (原编号100)
    if [(max(abs(x / 1500 - 1), abs(y / 175 - 1), abs(z / 12.5 - 1)) < 0.025)] != [
        (max(abs(x / 1500 - 1), abs(y / 175 - 1), abs(z / 12.5 - 1)) < 0.25)]:
        triggered.add(89)
    # 编号90 (原编号101)
    if [(min(x / 1500, y / 175, z / 12.5) > 0.98)] != [(min(x / 1500, y / 175, z / 12.5) > 498)]:
        triggered.add(90)
    # 编号91 (原编号102)
    if [(min(x / 1500, y / 175, z / 12.5) > 0.98)] != [(min(x / 1500, y / 1375, z / 12.5) > 0.98)]:
        triggered.add(91)
    # 编号92 (原编号103)
    if [(max(x / 1500, y / 175, z / 12.5) < 1.02)] != [(max(x / 1500, y / 175, z / 12.5) < 302)]:
        triggered.add(92)
    # 编号93 (原编号104)
    if [(max(x / 1500, y / 175, z / 12.5) < 1.02)] != [(max(x / 1500, y / 175, z / 12.5) < 102)]:
        triggered.add(93)
    # 编号94 (原编号105)
    if [(abs(max(x / 10, y, z * 10) - min(x / 10, y, z * 10)) < 30)] != [
        (abs(max(x / 10, y, z * 120) - min(x / 10, y, z * 10)) < 30)]:
        triggered.add(94)
    # 编号95 (原编号106)
    if [(abs(max(x / 10, y, z * 10) - min(x / 10, y, z * 10)) < 30)] != [
        (abs(max(x / 10, y, z * 10) - min(x / 10, y, z * 10)) < 230)]:
        triggered.add(95)
    # 编号96 (原编号107)
    if [((x / 10 + y + z * 10) > 378 and (x / 10 + y + z * 10) < 382)] != [
        ((x / 10 + y + z * 10) > 378 and (x / 10 + y + z * 10) < 3112)]:
        triggered.add(96)
    # 编号97 (原编号108)
    if [((x / 10 + y + z * 10) > 378 and (x / 10 + y + z * 10) < 382)] != [
        ((x / 10 + y + z * 10) > 3781 and (x / 10 + y + z * 10) < 382)]:
        triggered.add(97)
    # 编号98 (原编号109)
    if [((x * y * z / 10000) > 32.5 and (x * y * z / 10000) < 33.5)] != [
        ((x * y * z / 10000) > 32.5 and (x * y * 2 / 10000) < 33.5)]:
        triggered.add(98)
    # 编号99 (原编号110)
    if [((x * y * z / 10000) > 32.5 and (x * y * z / 10000) < 33.5)] != [
        ((x * y * z / 10000) > 32.5 and (x * y * 2.4 / 10000) < 33.5)]:
        triggered.add(99)
    # 编号100 (原编号111)
    if [(abs((x / 10 + y + z * 10) / 3 - 126.7) < 1.5)] != [(abs((x / 10 + y + z * 10) / 3 - 126.7) < 15)]:
        triggered.add(100)
    # 编号101 (原编号112)
    if [(abs((x / 10 + y + z * 10) / 3 - 126.7) < 1.5)] != [(abs((x / 10 + y + z * 10) / 3 - 126.7) < 115)]:
        triggered.add(101)
    # 编号102 (原编号113)
    if [(x > 1495 and x < 1505 and y > 174 and y < 176)] != [(x > 1495 and x < 1505 and y * 10 > 174 and y < 176)]:
        triggered.add(102)
    # 编号103 (原编号114)
    if [(x > 1495 and x < 1505 and y > 174 and y < 176)] != [(x > 1495 and x < 1505 and y * 88 > 174 and y < 176)]:
        triggered.add(103)
    # 编号104 (原编号115)
    if [(z > 12.45 and z < 12.55)] != [(z * 10 > 12.45 and z < 12.55)]:
        triggered.add(104)
    # 编号105 (原编号116)
    if [(z > 12.45 and z < 12.55)] != [(z * 8 > 12.45 and z < 12.55)]:
        triggered.add(105)

    # -------------------------- 106-144：动力控制优化 --------------------------
    # 编号106 (原编号117)
    if [(x < 1350 or y < 160 or z < 11.5)] != [(x < 1350 or y * 8 < 160 or z < 11.5)]:
        triggered.add(106)
    # 编号107 (原编号118)
    if [(x < 1350 or y < 160 or z < 11.5)] != [(x < 1350 or y < 160 or z * 89 < 11.5)]:
        triggered.add(107)
    # 编号108 (原编号119)
    if [(x > 1650 or y > 190 or z > 13.5)] != [(x > 1650 or y > 190 or z > 135)]:
        triggered.add(108)
    # 编号109 (原编号120)
    if [(x > 1650 or y > 190 or z > 13.5)] != [(x > 1650 or y * 87 > 190 or z > 13.5)]:
        triggered.add(109)
    # 编号110 (原编号121)
    if [(abs(x / (y * 10) - 0.86) > 0.1)] != [(abs(x / (y * 10) - 0.86) > 111)]:
        triggered.add(110)
    # 编号111 (原编号122)
    if [(abs(x / (y * 10) - 0.86) > 0.1)] != [(abs(x / (y * 10) - 0.86) > 1671)]:
        triggered.add(111)
    # 编号112 (原编号123)
    if [(abs(x / z - 120) > 15)] != [(abs(x / z - 120) > 1522)]:
        triggered.add(112)
    # 编号113 (原编号124)
    if [(abs(x / z - 120) > 15)] != [(abs(x / z - 120) > 115)]:
        triggered.add(113)
    # 编号114 (原编号125)
    if [(abs(y / z - 14) > 2)] != [(abs(y / z - 14) > 112)]:
        triggered.add(114)
    # 编号115 (原编号126)
    if [(abs(y / z - 14) > 2)] != [(abs(y / z - 14) > 211)]:
        triggered.add(115)
    # 编号116 (原编号127)
    if [((x / 10 + y + z * 10) < 360 or (x / 10 + y + z * 10) > 400)] != [
        ((x / 10 + y + z * 10) < 36 or (x / 10 + y + z * 10) > 400)]:
        triggered.add(116)
    # 编号117 (原编号128)
    if [((x / 10 + y + z * 10) < 360 or (x / 10 + y + z * 10) > 400)] != [
        ((x / 10 + y + z * 10) < 989 or (x / 10 + y + z * 10) > 400)]:
        triggered.add(117)
    # 编号118 (原编号129)
    if [((x * y * z / 10000) < 30 or (x * y * z / 10000) > 36)] != [
        ((x * y * z / 10000) < 30 or (x * y * z / 10000) > 136)]:
        triggered.add(118)
    # 编号119 (原编号130)
    if [((x * y * z / 10000) < 30 or (x * y * z / 10000) > 36)] != [
        ((x * y * z / 10000) < 30 or (x * y * z / 10000) > 316)]:
        triggered.add(119)
    # 编号120 (原编号131)
    if [(x < 1300 and y < 160)] != [(x < 1300 and y * 878 < 160)]:
        triggered.add(120)
    # 编号121 (原编号132)
    if [(x < 1300 and y < 160)] != [(x * 71 < 1300 and y < 160)]:
        triggered.add(121)
    # 编号122 (原编号133)
    if [(x > 1700 and y > 190)] != [(x > 1700 and y * 78 > 190)]:
        triggered.add(122)
    # 编号123 (原编号134)
    if [(x > 1700 and y > 190)] != [(x * 78 > 1700 and y > 190)]:
        triggered.add(123)
    # 编号124 (原编号135)
    if [(x < 1300 and z < 11)] != [(x < 1300 and z * 91 < 11)]:
        triggered.add(124)
    # 编号125 (原编号136)
    if [(x < 1300 and z < 11)] != [(x * 12 < 1300 and z < 11)]:
        triggered.add(125)
    # 编号126 (原编号137)
    if [(x > 1700 and z > 14)] != [(x > 1700 and z * 21 > 14)]:
        triggered.add(126)
    # 编号127 (原编号138)
    if [(x > 1700 and z > 14)] != [(x * 123 > 1700 and z > 14)]:
        triggered.add(127)
    # 编号128 (原编号139)
    if [(y < 160 and z < 11)] != [(y < 160 and z * 78 < 11)]:
        triggered.add(128)
    # 编号129 (原编号140)
    if [(y < 160 and z < 11)] != [(y * 8 < 160 and z < 11)]:
        triggered.add(129)
    # 编号130 (原编号141)
    if [(y > 190 and z > 14)] != [(y * 8 > 190 and z > 14)]:
        triggered.add(130)
    # 编号131 (原编号142)
    if [(y > 190 and z > 14)] != [(y > 190 and z * 9 > 14)]:
        triggered.add(131)
    # 编号132 (原编号143)
    if [(x < 1250 or x > 1750)] != [(x * 67 < 1250 or x > 1750)]:
        triggered.add(132)
    # 编号133 (原编号144)
    if [(x < 1250 or x > 1750)] != [(x < 1250 or x * 53 > 1750)]:
        triggered.add(133)
    # 编号134 (原编号145)
    if [(y < 145 or y > 205)] != [(y * 67 < 145 or y > 205)]:
        triggered.add(134)
    # 编号135 (原编号146)
    if [(y < 145 or y > 205)] != [(y < 145 or y * 67 > 205)]:
        triggered.add(135)
    # 编号136 (原编号147)
    if [(z < 9.5 or z > 15.5)] != [(z < 9.5 or z * 67 > 15.5)]:
        triggered.add(136)
    # 编号137 (原编号149)
    if [(x < 1200 and y < 155 and z < 10.5)] != [(x * 67 < 1200 and y < 155 and z < 10.5)]:
        triggered.add(137)
    # 编号138 (原编号150)
    if [(x < 1200 and y < 155 and z < 10.5)] != [(x < 1200 and y < 55 and z < 10.5)]:
        triggered.add(138)
    # 编号139 (原编号151)
    if [(x > 1800 and y > 195 and z > 14.5)] != [(x > 1800 and y * 78 > 195 and z > 14.5)]:
        triggered.add(139)
    # 编号140 (原编号152)
    if [(x > 1800 and y > 195 and z > 14.5)] != [(x > 1800 and y > 966 and z > 14.5)]:
        triggered.add(140)
    # 编号141 (原编号153)
    if [((x / 10 - 150) ** 2 + (y - 175) ** 2 + (z - 12.5) ** 2 * 100) > 150] != [
        ((x / 10 - 150) ** 6 + (y - 175) ** 2 + (z - 12.5) ** 2 * 100) > 150]:
        triggered.add(141)
    # 编号142 (原编号154)
    if [((x / 10 - 150) ** 2 + (y - 175) ** 2 + (z - 12.5) ** 2 * 100) > 150] != [
        ((x / 10 - 150) ** 2 + (y - 175) ** 8 + (z - 12.5) ** 2 * 100) > 150]:
        triggered.add(142)
    # 编号143 (原编号155)
    if [(abs((x * y * z / 10000) / 33 - 1) > 0.18)] != [(abs((x * y * z / 10000) / 33 - 1) > 18)]:
        triggered.add(143)
    # 编号144 (原编号156)
    if [(abs((x * y * z / 10000) / 33 - 1) > 0.18)] != [(abs((x * y * z / 10000) / 33 - 1) > 10)]:
        triggered.add(144)

    return triggered


def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """"""
    generated_path = execute_validation_rules(particle[0], particle[1], particle[2])

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # : x:1-50, y:1-50, z:1-50
        self.bounds = bounds if bounds else [(1000, 2000), (100, 200), (10, 30)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """target pathsPSO"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(particles[i][0], particles[i][1], particles[i][2]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(gbest_particle[0], gbest_particle[1], gbest_particle[2]),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """Path PSO"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - Path ")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations")
    print(f"Path : {len(target_paths)}")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"Path {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "" if result['success'] else f"({result['best_fitness']:.3f})"
        print(f"{status} | {result['time']:.2f}s | iterations{result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f": {success_count}/{len(target_paths)} ({success_rate:.1f}%) | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs}")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations, {len(target_paths)}Path ")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"---  {run_idx}/{num_runs} ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f": {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs} runcompleted | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 1: 
    ws1 = wb.active
    ws1.title = ""
    ws1.sheet_view.showGridLines = False

    headers = ["Run", "", "", "", "Average Iterations", "(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f" {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # 2: Path 
    ws2 = wb.create_sheet(title="Path ")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Path ID", "", "", "", "Average Iterations", "Minimum Iterations", "Maximum Iterations"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"Path  {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # 3: 
    ws3 = wb.create_sheet(title="")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Path ", "", "(x,y,z)", "", "Iterations", "Path "]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"{run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # 4: target paths
    ws4 = wb.create_sheet(title="target paths")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Path ID", "target paths", ""]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"Path  {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f" : {filename}")
    print(f"{'=' * 70}")
    print(f":")
    print(f"  1.        - {len(all_results)} run")
    print(f"  2. Path        - Path ")
    print(f"  3.    -  runPath ")
    print(f"  4. target paths       - target paths")
    print(f"{'=' * 70}\n")

    return filename


def main():


    target_paths = [
        {1, 4, 6, 9, 11, 12, 13, 14, 15, 16, 17, 21, 22, 24, 25, 27, 29, 32, 33, 38, 44, 45, 46, 47, 48, 49, 51, 52, 58,
         59,
         60, 64, 68, 69, 74, 75, 78, 79, 88, 89, 92, 93, 95, 96, 101, 104, 105, 107, 109, 110, 111, 112, 113, 114, 115,
         119,
         133, 135, 136, 143, 144},

        {4, 6, 9, 11, 12, 13, 14, 15, 16, 17, 21, 22, 24, 25, 27, 29, 32, 34, 37, 39, 40, 44, 45, 47, 48, 50, 51, 52,
         58,
         59, 60, 61, 64, 68, 69, 74, 75, 78, 79, 88, 89, 92, 93, 95, 96, 101, 104, 105, 109, 110, 111, 112, 113, 119,
         128,
         129, 133, 135, 136, 143, 144},

        {3, 5, 9, 11, 12, 13, 14, 16, 17, 18, 21, 22, 24, 25, 27, 29, 32, 34, 35, 36, 37, 39, 40, 44, 45, 49, 51, 52,
         54,
         58, 59, 60, 61, 65, 74, 75, 78, 79, 86, 87, 88, 89, 92, 93, 95, 96, 101, 104, 105, 106, 109, 110, 111, 112,
         113,
         119, 133, 135, 136, 143, 144},

        {2, 3, 8, 9, 11, 12, 13, 14, 16, 17, 18, 19, 21, 22, 23, 27, 29, 31, 34, 35, 36, 41, 42, 44, 45, 49, 51, 54, 56,
         57,
         62, 64, 66, 67, 68, 69, 78, 79, 88, 89, 94, 96, 100, 101, 104, 105, 106, 109, 110, 111, 114, 115, 117, 119,
         133,
         134, 136, 143, 144},

        {2, 4, 7, 9, 11, 12, 16, 17, 21, 22, 24, 27, 29, 32, 33, 35, 36, 37, 41, 42, 44, 45, 50, 51, 52, 57, 64, 66, 67,
         68,
         69, 78, 79, 83, 88, 89, 95, 96, 100, 101, 102, 103, 104, 105, 109, 110, 111, 112, 113, 119, 128, 129, 133, 135,
         136, 143, 144},

        {6, 8, 11, 12, 14, 15, 16, 17, 21, 22, 24, 29, 32, 35, 36, 38, 39, 40, 42, 43, 45, 49, 52, 54, 56, 57, 60, 61,
         63,
         65, 66, 67, 68, 69, 88, 89, 92, 93, 95, 96, 100, 101, 104, 105, 109, 110, 111, 114, 115, 119, 124, 125, 132,
         135,
         136, 143, 144},

        {1, 4, 6, 8, 9, 11, 12, 13, 14, 15, 16, 17, 21, 22, 24, 25, 29, 32, 33, 35, 36, 37, 39, 40, 42, 44, 49, 52, 54,
         57,
         59, 60, 65, 66, 67, 68, 69, 72, 73, 78, 79, 88, 89, 95, 96, 100, 101, 104, 105, 107, 109, 119, 133, 135, 136,
         143,
         144},

        {1, 3, 5, 7, 9, 11, 12, 17, 18, 21, 22, 24, 25, 27, 30, 32, 33, 35, 36, 37, 39, 40, 41, 42, 44, 50, 52, 54, 58,
         59,
         65, 72, 73, 74, 75, 78, 79, 80, 81, 82, 88, 89, 95, 96, 101, 102, 103, 104, 105, 109, 119, 133, 135, 136, 142,
         143,
         144},

        {8, 9, 11, 12, 14, 16, 17, 21, 22, 23, 29, 31, 33, 35, 36, 37, 44, 46, 49, 54, 56, 62, 63, 64, 66, 67, 68, 69,
         72,
         73, 78, 79, 88, 89, 95, 98, 100, 101, 104, 105, 109, 117, 119, 120, 121, 124, 125, 128, 129, 132, 134, 136,
         137,
         138, 143, 144},

        {9, 11, 12, 14, 16, 17, 24, 25, 27, 29, 32, 34, 37, 39, 40, 45, 47, 48, 50, 51, 52, 57, 59, 60, 61, 64, 68, 69,
         74,
         75, 78, 79, 88, 89, 92, 93, 95, 96, 100, 101, 104, 105, 110, 111, 112, 113, 119, 122, 126, 128, 129, 133, 135,
         136,
         143, 144},

        {6, 8, 11, 12, 14, 16, 17, 21, 22, 23, 29, 31, 35, 36, 38, 39, 40, 43, 45, 49, 52, 54, 56, 57, 60, 61, 63, 65,
         66,
         67, 68, 69, 88, 89, 92, 93, 95, 96, 100, 101, 104, 105, 110, 111, 114, 115, 119, 123, 124, 125, 131, 132, 135,
         136,
         143, 144},

        {8, 9, 11, 12, 14, 16, 17, 21, 22, 29, 34, 35, 36, 42, 44, 45, 49, 51, 54, 56, 62, 63, 64, 66, 67, 68, 69, 78,
         79,
         88, 94, 98, 99, 100, 101, 104, 105, 109, 110, 111, 114, 115, 116, 119, 120, 121, 124, 125, 128, 129, 133, 134,
         136,
         143, 144},

        {5, 8, 11, 12, 14, 16, 17, 20, 21, 22, 24, 30, 32, 38, 39, 40, 42, 43, 45, 49, 52, 54, 58, 60, 61, 63, 65, 66,
         67,
         68, 69, 88, 89, 92, 93, 95, 96, 100, 101, 104, 105, 110, 111, 112, 113, 114, 115, 119, 123, 131, 132, 135, 136,
         143, 144},

        {8, 9, 14, 16, 17, 23, 27, 29, 31, 34, 41, 42, 44, 45, 47, 48, 49, 51, 55, 56, 57, 62, 66, 67, 68, 69, 84, 85,
         88,
         92, 93, 95, 96, 98, 99, 100, 101, 108, 110, 111, 112, 113, 114, 115, 117, 119, 120, 121, 127, 130, 132, 134,
         136,
         143, 144},

        {5, 8, 9, 14, 15, 16, 17, 18, 20, 21, 22, 23, 30, 32, 37, 41, 42, 44, 45, 46, 47, 48, 49, 52, 55, 58, 62, 65,
         66,
         67, 68, 69, 76, 77, 84, 85, 88, 89, 92, 93, 95, 96, 100, 101, 109, 110, 111, 112, 113, 119, 132, 135, 136, 143,
         144},

        {1, 3, 5, 9, 11, 12, 17, 18, 21, 22, 24, 25, 27, 30, 32, 33, 35, 36, 37, 39, 40, 44, 50, 52, 54, 58, 59, 60, 65,
         72,
         73, 74, 75, 78, 79, 80, 81, 82, 88, 89, 95, 96, 101, 104, 105, 109, 119, 133, 135, 136, 141, 142, 143, 144},

        {1, 4, 6, 8, 10, 11, 12, 17, 21, 22, 24, 25, 29, 32, 33, 35, 36, 38, 39, 40, 42, 44, 49, 52, 58, 59, 68, 69, 72,
         73,
         74, 75, 80, 81, 82, 88, 89, 95, 96, 101, 104, 105, 107, 109, 112, 113, 114, 115, 119, 133, 135, 136, 143, 144},

        {8, 9, 11, 12, 14, 16, 17, 21, 22, 23, 29, 31, 33, 35, 36, 37, 42, 44, 46, 50, 52, 54, 56, 62, 63, 64, 66, 67,
         68,
         69, 72, 73, 78, 79, 88, 89, 95, 97, 104, 105, 109, 117, 119, 120, 121, 124, 125, 128, 129, 133, 135, 136, 143,
         144},

        {8, 9, 11, 12, 14, 16, 17, 21, 22, 29, 34, 42, 44, 45, 49, 51, 54, 56, 63, 64, 66, 67, 68, 69, 78, 79, 88, 94,
         98,
         99, 101, 104, 105, 109, 110, 111, 112, 113, 114, 115, 116, 118, 119, 120, 121, 124, 125, 128, 129, 132, 134,
         136,
         143, 144},

        {5, 11, 12, 14, 15, 16, 17, 24, 26, 27, 30, 32, 33, 38, 43, 45, 46, 47, 48, 49, 51, 52, 58, 59, 62, 66, 67, 68,
         69,
         78, 79, 88, 89, 92, 93, 95, 96, 101, 104, 105, 107, 110, 111, 112, 113, 114, 115, 122, 126, 135, 136, 143},

        {1, 3, 5, 7, 9, 21, 22, 24, 25, 27, 30, 32, 33, 35, 36, 37, 39, 40, 44, 50, 52, 54, 58, 59, 65, 70, 71, 72, 73,
         80,
         81, 82, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 95, 96, 101, 109, 133, 135, 136, 142, 143, 144},

        {2, 8, 9, 13, 14, 16, 17, 19, 24, 28, 30, 32, 34, 44, 45, 47, 48, 49, 51, 55, 58, 59, 60, 62, 76, 77, 84, 85,
         86,
         87, 88, 92, 93, 94, 96, 101, 106, 108, 110, 111, 112, 113, 114, 115, 119, 127, 130, 133, 134, 143, 144},

        {14, 15, 16, 17, 24, 26, 28, 32, 33, 35, 36, 37, 43, 45, 46, 49, 51, 52, 54, 59, 62, 63, 64, 65, 66, 67, 68, 69,
         84,
         85, 88, 89, 90, 91, 92, 93, 95, 96, 101, 110, 111, 122, 130, 135, 136, 139, 143},

        {11, 12, 14, 16, 17, 24, 26, 27, 29, 32, 33, 38, 43, 46, 47, 48, 49, 51, 53, 58, 59, 62, 65, 72, 73, 78, 79, 88,
         89,
         92, 93, 95, 96, 101, 104, 105, 107, 112, 113, 114, 115, 126, 131, 135, 136, 143, 144},

        {14, 16, 17, 24, 26, 28, 32, 33, 35, 36, 37, 45, 49, 51, 52, 54, 59, 62, 63, 64, 65, 66, 67, 68, 69, 84, 85, 88,
         90,
         91, 92, 93, 95, 96, 101, 110, 111, 135, 136, 140, 143},
    ]

    print("=" * 70)
    print("baseline PSO")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("Program completed")


if __name__ == "__main__":
    main()