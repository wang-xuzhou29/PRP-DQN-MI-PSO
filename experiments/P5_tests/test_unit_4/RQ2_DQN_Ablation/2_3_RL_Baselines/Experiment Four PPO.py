
import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
import numpy as np
import random
import time
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
from datetime import datetime

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# ===  ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'ACTION_DIM': 3,
    'MIN_VALUES': np.array([1, 1, 1]),
    'MAX_VALUES': np.array([128, 200, 255]),
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 3,
    'PPO_BATCH_SIZE': 64,
    'PPO_EPOCHS': 4,
    'CLIP_EPSILON': 0.2,
    'GAMMA': 0.99,
    'GAE_LAMBDA': 0.95,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'ACTOR_LR': 3e-4,
    'CRITIC_LR': 3e-4,
    'NUM_RUNS': 20,
    'TOP_K_SAMPLES': 20,
    'TARGET_PATHS': [
        {15, 16, 48, 49, 74, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 91, 92, 93, 94, 99, 100},
        {16, 18, 19, 60, 61, 70, 71, 72, 73, 80, 81, 82, 83, 89, 91, 92, 93, 94, 99, 100},
        {1, 4, 6, 46, 47, 74, 75, 76, 77, 78, 80, 81, 82, 83, 89, 92, 93, 94, 99, 100},
        {30, 31, 50, 51, 70, 71, 72, 73, 84, 85, 86, 87, 88, 91, 92, 93, 95, 99, 100},
        {18, 19, 36, 37, 74, 76, 77, 78, 79, 80, 81, 82, 83, 89, 92, 93, 94, 99, 100},
        {20, 24, 25, 36, 37, 76, 77, 78, 84, 86, 87, 88, 90, 91, 92, 93, 95, 99, 100},
        {8, 12, 34, 35, 74, 75, 76, 77, 78, 84, 86, 87, 88, 90, 91, 92, 93, 95, 100},
        {8, 10, 58, 59, 70, 71, 72, 73, 84, 85, 86, 87, 88, 91, 92, 93, 95, 99, 100},
        {8, 14, 46, 47, 75, 76, 77, 78, 84, 85, 86, 87, 88, 90, 92, 93, 95, 99, 100},
        {1, 2, 6, 46, 47, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 92, 93, 94, 100},
        {39, 74, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 91, 92, 93, 94, 99, 100},
        {20, 21, 60, 61, 70, 71, 72, 73, 84, 85, 86, 87, 88, 90, 92, 93, 95, 99},
        {8, 9, 11, 13, 40, 41, 75, 76, 77, 78, 79, 80, 81, 83, 96, 97, 98, 100},
        {18, 19, 54, 55, 70, 71, 72, 73, 84, 86, 87, 88, 96, 97, 98, 99, 100},
        {27, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 91, 92, 93, 94, 99, 100},
        {25, 48, 49, 69, 71, 72, 73, 84, 85, 86, 87, 88, 90, 92, 93, 95, 100},
        {26, 28, 62, 70, 71, 72, 73, 80, 81, 82, 83, 89, 91, 92, 93, 94, 100},
        {32, 33, 68, 69, 71, 72, 73, 79, 80, 81, 82, 83, 96, 97, 98, 99, 100},
        {1, 52, 53, 74, 75, 76, 77, 78, 84, 85, 86, 87, 88, 97, 98, 99, 100},
        {8, 12, 14, 64, 65, 69, 71, 72, 73, 80, 81, 82, 83, 96, 97, 98, 100},
        {1, 3, 64, 65, 70, 71, 72, 73, 84, 86, 87, 88, 96, 97, 98, 99, 100},
        {22, 36, 37, 76, 77, 78, 85, 86, 87, 88, 90, 91, 93, 95, 100},
        {31, 45, 70, 71, 72, 73, 79, 80, 81, 83, 96, 97, 98, 99, 100},
        {22, 66, 67, 69, 71, 72, 73, 79, 80, 82, 83, 97, 98, 100},
        {44, 45, 69, 71, 72, 73, 79, 80, 83, 96, 97, 98, 99, 100},
        {57, 71, 72, 73, 79, 80, 83, 97, 98, 100},
        {15, 16, 17, 48, 49, 74, 75, 76, 77, 78, 79, 80, 82, 83, 89, 91, 92, 93, 94, 100},
        {1, 2, 5, 46, 47, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 91, 92, 93, 94, 100},
        {20, 21, 25, 42, 43, 74, 76, 77, 78, 79, 80, 81, 82, 83, 96, 97, 98, 99, 100},
        {2, 5, 7, 40, 41, 75, 76, 77, 78, 84, 85, 86, 87, 88, 96, 97, 98, 99, 100},
        {26, 28, 56, 57, 70, 71, 72, 73, 84, 85, 86, 87, 88, 96, 97, 98, 99, 100},
        {26, 28, 38, 74, 76, 77, 78, 80, 81, 82, 83, 89, 91, 92, 93, 94, 100},
        {30, 31, 62, 63, 70, 71, 72, 73, 84, 86, 87, 88, 90, 91, 92, 93, 95},
        {29, 62, 63, 71, 72, 73, 84, 85, 86, 87, 88, 90, 92, 93, 95, 100},
        {23, 25, 60, 61, 71, 72, 73, 84, 85, 86, 87, 88, 90, 92, 93, 95, 100}
    ],
}

# ===  ===
def execute_Tr(weather, time_period, z):
    """验证规则函数 - weather, time_period, z组合（删除覆盖率为0的分支后重新编号版本）"""

    triggered = set()

    # 将z映射到合适的范围以匹配条件逻辑
    x = z  # 直接使用z作为x
    y = (weather * time_period * 10 + z) % 100 + 1  # 基于输入参数计算y值

    # 1-7: 早高峰组合（time_period == 1）
    if time_period == 1:
        if (x < 60 and y > 75) != (x < 60 and y < 75):
            triggered.add(1)
        if (x > 60 and y > 70) != (x < 60 and y > 70):
            triggered.add(2)
        if (x < 50 and y < 40) != (x > 50 and y < 40):
            triggered.add(3)
        if (x > 78 and 45 < y < 70) != (x < 78 and 45 < y < 70):
            triggered.add(4)
        if (45 < x < 70 and y > 78) != (45 < x < 70 and y < 78):
            triggered.add(5)
        if (x < 55 and 50 < y < 75) != (x > 55 and 50 < y < 75):
            triggered.add(6)
        if (50 < x < 75 and y < 55) != (50 < x < 75 and y > 55):
            triggered.add(7)

    # 8-14: 晚高峰组合（time_period == 2）
    if time_period == 2:
        if (x < 60 and y > 75) != (x < 60 and y < 75):
            triggered.add(8)
        if (x > 60 and y > 70) != (x < 60 and y > 70):
            triggered.add(9)
        if (x < 55 and y < 45) != (x > 55 and y < 45):
            triggered.add(10)
        if (45 < x < 70 and y > 78) != (45 < x < 70 and y < 78):
            triggered.add(11)
        if (x > 78 and 45 < y < 70) != (x < 78 and 45 < y < 70):
            triggered.add(12)
        if (55 < x < 75 and y < 50) != (55 < x < 75 and y > 50):
            triggered.add(13)
        if (x < 50 and 55 < y < 75) != (x > 50 and 55 < y < 75):
            triggered.add(14)

    # 15-19: 午餐时间组合（time_period == 3）
    if time_period == 3:
        if (x > 60 and 40 < y < 65) != (x < 60 and 40 < y < 65):
            triggered.add(15)
        if (40 < x < 65 and y > 60) != (40 < x < 65 and y < 60):
            triggered.add(16)
        if (45 < x < 70 and 45 < y < 60) != (45 < x < 70 and y > 60):
            triggered.add(17)  # Fixed: 45 > y > 60 was invalid
        if (x < 50 and y < 40) != (x > 50 and y < 40):
            triggered.add(18)
        if (x > 65 and y < 45) != (x < 65 and y < 45):
            triggered.add(19)

    # 20-25: 夜间组合（time_period == 4）
    if time_period == 4:
        if (x < 45 and y < 35) != (x > 45 and y < 35):
            triggered.add(20)
        if (x > 60 and y < 40) != (x < 60 and y < 40):
            triggered.add(21)
        if (x < 50 and y > 70) != (x > 50 and y > 70):
            triggered.add(22)
        if (45 < x < 70 and 45 < y < 60) != (x > 70 and 45 < y < 60):
            triggered.add(23)  # Fixed: 45 > x > 70 was invalid
        if (x < 35 and y < 25) != (x > 35 and y < 25):
            triggered.add(24)
        if (40 < x < 65 and y < 45) != (40 < x < 65 and y > 45):
            triggered.add(25)

    # 26-28: 周末组合（time_period == 5）
    if time_period == 5:
        if (x < 60 and y < 50) != (x > 60 and y < 50):
            triggered.add(26)
        if (x > 65 and y > 75) != (x < 65 and y > 75):
            triggered.add(27)
        if (x > 60 and y < 45) != (x < 60 and y < 45):
            triggered.add(28)

    # 29-33: 假日组合（time_period == 6）
    if time_period == 6:
        if (40 < x < 70 and 40 < y < 60) != (x > 70 and 40 < y < 60):
            triggered.add(29)  # Fixed: 40 > x > 70 was invalid
        if (x < 55 and y < 45) != (x > 55 and y < 45):
            triggered.add(30)
        if (x > 60 and y < 50) != (x < 60 and y < 50):
            triggered.add(31)
        if (x < 60 and y > 70) != (x > 60 and y > 70):
            triggered.add(32)
        if (x > 65 and y > 75) != (x < 65 and y > 75):
            triggered.add(33)

    # 34-68: 天气相关扩展规则
    if weather == 1:  # 晴天
        if (time_period in [1, 2] and x > 70) != (time_period in [1, 2] and x < 70):
            triggered.add(34)
        if (time_period in [1, 2] and y > 70) != (time_period in [1, 2] and y < 70):
            triggered.add(35)
        if (time_period in [3, 4] and x < 50) != (time_period in [3, 4] and x > 50):
            triggered.add(36)
        if (time_period in [3, 4] and y < 50) != (time_period in [3, 4] and y > 50):
            triggered.add(37)
        if (time_period in [5, 6] and 40 < x < 80) != (time_period in [5, 6] and x > 80):
            triggered.add(38)  # Fixed: 40 > x > 80 was invalid
        if (time_period in [5, 6] and 40 < y < 80) != (time_period in [5, 6] and y > 80):
            triggered.add(39)  # Fixed: 40 > y > 80 was invalid

    if weather == 2:  # 雨天
        if (time_period in [1, 2] and x > 75) != (time_period in [1, 2] and x < 75):
            triggered.add(40)
        if (time_period in [1, 2] and y < 60) != (time_period in [1, 2] and y > 60):
            triggered.add(41)
        if (time_period in [3, 4] and x < 45) != (time_period in [3, 4] and x > 45):
            triggered.add(42)
        if (time_period in [3, 4] and y > 65) != (time_period in [3, 4] and y < 65):
            triggered.add(43)
        if (time_period in [5, 6] and 35 < x < 75) != (time_period in [5, 6] and x > 75):
            triggered.add(44)  # Fixed: 35 > x > 75 was invalid
        if (time_period in [5, 6] and 35 < y < 75) != (time_period in [5, 6] and y > 75):
            triggered.add(45)  # Fixed: 35 > y > 75 was invalid

    if weather == 3:  # 雾天
        if (time_period in [1, 2] and x > 60) != (time_period in [1, 2] and x < 60):
            triggered.add(46)
        if (time_period in [1, 2] and y > 65) != (time_period in [1, 2] and y < 65):
            triggered.add(47)
        if (time_period in [3, 4] and x < 55) != (time_period in [3, 4] and x > 55):
            triggered.add(48)
        if (time_period in [3, 4] and y < 55) != (time_period in [3, 4] and y > 55):
            triggered.add(49)
        if (time_period in [5, 6] and 30 < x < 70) != (time_period in [5, 6] and x > 70):
            triggered.add(50)  # Fixed: 30 > x > 70 was invalid
        if (time_period in [5, 6] and 30 < y < 70) != (time_period in [5, 6] and y > 70):
            triggered.add(51)  # Fixed: 30 > y > 70 was invalid

    if weather == 4:  # 雪天
        if (time_period in [1, 2] and x > 65) != (time_period in [1, 2] and x < 65):
            triggered.add(52)
        if (time_period in [1, 2] and y < 55) != (time_period in [1, 2] and y > 55):
            triggered.add(53)
        if (time_period in [3, 4] and x < 40) != (time_period in [3, 4] and x > 40):
            triggered.add(54)
        if (time_period in [3, 4] and y > 60) != (time_period in [3, 4] and y < 60):
            triggered.add(55)
        if (time_period in [5, 6] and 25 < x < 65) != (time_period in [5, 6] and x > 65):
            triggered.add(56)  # Fixed: 25 > x > 65 was invalid
        if (time_period in [5, 6] and 25 < y < 65) != (time_period in [5, 6] and y > 65):
            triggered.add(57)  # Fixed: 25 > y > 65 was invalid

    if weather == 5:  # 风天
        if (time_period in [1, 2] and x > 70) != (time_period in [1, 2] and x < 70):
            triggered.add(58)
        if (time_period in [1, 2] and y > 60) != (time_period in [1, 2] and y < 60):
            triggered.add(59)
        if (time_period in [3, 4] and x < 35) != (time_period in [3, 4] and x > 35):
            triggered.add(60)
        if (time_period in [3, 4] and y < 40) != (time_period in [3, 4] and y > 40):
            triggered.add(61)
        if (time_period in [5, 6] and 20 < x < 60) != (time_period in [5, 6] and x > 60):
            triggered.add(62)  # Fixed: 20 > x > 60 was invalid
        if (time_period in [5, 6] and 20 < y < 60) != (time_period in [5, 6] and y > 60):
            triggered.add(63)  # Fixed: 20 > y > 60 was invalid

    if weather == 6:  # 暴雨
        if (time_period in [1, 2] and x > 55) != (time_period in [1, 2] and x < 55):
            triggered.add(64)
        if (time_period in [1, 2] and y > 55) != (time_period in [1, 2] and y < 55):
            triggered.add(65)
        if (time_period in [3, 4] and x < 45) != (time_period in [3, 4] and x > 45):
            triggered.add(66)
        if (time_period in [3, 4] and y < 45) != (time_period in [3, 4] and y > 45):
            triggered.add(67)
        if (time_period in [5, 6] and 15 < x < 55) != (time_period in [5, 6] and x > 55):
            triggered.add(68)  # Fixed: 15 > x > 55 was invalid

    # 69-78: 复合条件（多参数组合）
    if weather + time_period > 6:
        if (x > 50 and y > 50) != (x < 50 and y > 50):
            triggered.add(69)
        if (x < 50 and y < 50) != (x > 50 and y < 50):
            triggered.add(70)
        if (x > y) != (x < y):
            triggered.add(71)
        if (x < y) != (x > y):
            triggered.add(72)
        if (abs(x - y) < 20) != (abs(x - y) > 20):
            triggered.add(73)

    if weather + time_period <= 6:
        if (x > 60 or y > 60) != (x < 60 or y > 60):
            triggered.add(74)
        if (x < 40 or y < 40) != (x > 40 or y < 40):
            triggered.add(75)
        if (x + y > 100) != (x + y < 100):
            triggered.add(76)
        if (x + y < 80) != (x + y > 80):
            triggered.add(77)
        if (abs(x - y) > 30) != (abs(x - y) < 30):
            triggered.add(78)

    # 79-88: 数值关系条件
    if weather % 2 == time_period % 2:  # 同奇偶性
        if (x % 10 < 5) != (x % 10 > 5):
            triggered.add(79)
        if (y % 10 >= 5) != (y % 10 < 5):
            triggered.add(80)
        if ((x + y) % 3 == 0) != ((x + y) % 3 == 1):
            triggered.add(81)
        if ((x * y) % 7 == 0) != ((x * y) % 7 == 1):
            triggered.add(82)
        if (x // 10 == y // 10) != (x // 10 != y // 10):
            triggered.add(83)

    if weather % 2 != time_period % 2:  # 不同奇偶性
        if (x > 75 or y > 75) != (x < 75 or y > 75):
            triggered.add(84)
        if (x < 25 or y < 25) != (x > 25 or y < 25):
            triggered.add(85)
        if (max(x, y) - min(x, y) > 40) != (max(x, y) - min(x, y) < 40):
            triggered.add(86)
        if ((x + y) // 2 > 50) != ((x + y) // 2 < 50):
            triggered.add(87)
        if (weather * time_period > 15) != (weather * time_period < 15):
            triggered.add(88)

    # 89-95: 高级组合条件（奇数天气）
    if weather in [1, 3, 5]:  # 奇数天气
        if (time_period in [1, 3, 5] and x > 40) != (time_period in [1, 3, 5] and x < 40):
            triggered.add(89)
        if (time_period in [2, 4, 6] and y > 40) != (time_period in [2, 4, 6] and y < 40):
            triggered.add(90)
        if (x % 20 < 10 and y % 20 < 10) != (x % 20 > 10 and y % 20 < 10):
            triggered.add(91)
        if (x + weather * 10 > 50) != (x + weather * 10 < 50):
            triggered.add(92)
        if (y + time_period * 10 > 50) != (y + time_period * 10 < 50):
            triggered.add(93)
        if (time_period in [1, 3, 5] and x < 60) != (time_period in [1, 3, 5] and x > 60):
            triggered.add(94)
        if (time_period in [2, 4, 6] and y < 60) != (time_period in [2, 4, 6] and y > 60):
            triggered.add(95)

    # 96-98: 偶数天气条件
    if weather in [2, 4, 6]:  # 偶数天气
        if ((x + y) % weather == 0) != ((x + y) % weather == 1):
            triggered.add(96)
        if (x * weather > 100) != (x * weather < 100):
            triggered.add(97)
        if (y * time_period > 100) != (y * time_period < 100):
            triggered.add(98)

    # 99-100: 最后的复杂条件
    if ((weather * time_period + z) % 7 == 0) != ((weather * time_period + z) % 7 == 1):
        triggered.add(99)
    if (max(weather, time_period) * min(x, y) > 150) != (max(weather, time_period) * min(x, y) < 150):
        triggered.add(100)

    return triggered
# ===  ===
# === 状态处理辅助函数 ===
def clip_state(state):
    return np.clip(state, EXPERIMENT_CONFIG['MIN_VALUES'], EXPERIMENT_CONFIG['MAX_VALUES'])

def denormalize_state(normalized_state):
    """将归一化状态还原为原始状态"""
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return normalized_state * (max_vals - min_vals) / 2 + (min_vals + max_vals) / 2

def coverage_similarity(triggered, target_path):
    """
    Similarity:  / target paths
    
    """
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0

    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)

def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']

    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']

    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']

    return reward

# === PPO Actor ===
class PPOActor(nn.Module):
    def __init__(self, state_dim=3, action_dim=3, hidden_dim=256):
        super(PPOActor, self).__init__()

        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)

        self.mean_head = nn.Linear(hidden_dim, action_dim)
        self.log_std_head = nn.Linear(hidden_dim, action_dim)

        self.action_scale = 10.0

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))
        x = F.relu(self.fc3(x))

        mean = torch.tanh(self.mean_head(x)) * self.action_scale
        log_std = torch.clamp(self.log_std_head(x), -20, 2)

        return mean, log_std

    def get_action_and_log_prob(self, state):
        mean, log_std = self.forward(state)
        std = torch.exp(log_std)

        dist = torch.distributions.Normal(mean, std)
        action = dist.sample()
        log_prob = dist.log_prob(action).sum(dim=-1)

        return action, log_prob

    def get_log_prob(self, state, action):
        mean, log_std = self.forward(state)
        std = torch.exp(log_std)

        dist = torch.distributions.Normal(mean, std)
        log_prob = dist.log_prob(action).sum(dim=-1)

        return log_prob

# === PPO Critic ===
class PPOCritic(nn.Module):
    def __init__(self, state_dim=3, hidden_dim=256):
        super(PPOCritic, self).__init__()

        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)
        self.value_head = nn.Linear(hidden_dim, 1)

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))
        x = F.relu(self.fc3(x))
        value = self.value_head(x)
        return value.squeeze(-1)

# === PPO ===
class PPOBuffer:
    def __init__(self):
        self.states = []
        self.actions = []
        self.rewards = []
        self.values = []
        self.log_probs = []
        self.dones = []
        self.path_indices = []
        self.similarities = []

    def store(self, state, action, reward, value, log_prob, done, path_idx, similarity):
        self.states.append(state)
        self.actions.append(action)
        self.rewards.append(reward)
        self.values.append(value)
        self.log_probs.append(log_prob)
        self.dones.append(done)
        self.path_indices.append(path_idx)
        self.similarities.append(similarity)

    def compute_advantages(self):
        rewards = np.array(self.rewards)
        values = np.array(self.values)
        dones = np.array(self.dones)

        advantages = np.zeros_like(rewards)
        returns = np.zeros_like(rewards)
        last_advantage = 0

        for t in reversed(range(len(rewards))):
            if t == len(rewards) - 1:
                next_value = 0
            else:
                next_value = values[t + 1]

            delta = rewards[t] + EXPERIMENT_CONFIG['GAMMA'] * next_value * (1 - dones[t]) - values[t]
            advantages[t] = delta + EXPERIMENT_CONFIG['GAMMA'] * EXPERIMENT_CONFIG['GAE_LAMBDA'] * (
                    1 - dones[t]) * last_advantage
            last_advantage = advantages[t]

        returns = advantages + values
        advantages = (advantages - np.mean(advantages)) / (np.std(advantages) + 1e-8)

        return advantages, returns

    def get_batch_iterator(self, advantages, returns):
        indices = np.arange(len(self.states))
        np.random.shuffle(indices)

        batch_size = EXPERIMENT_CONFIG['PPO_BATCH_SIZE']
        for start in range(0, len(self.states), batch_size):
            end = min(start + batch_size, len(self.states))
            batch_indices = indices[start:end]

            yield {
                'states': torch.FloatTensor(np.array([self.states[i] for i in batch_indices])).to(device),
                'actions': torch.FloatTensor(np.array([self.actions[i] for i in batch_indices])).to(device),
                'old_log_probs': torch.FloatTensor(np.array([self.log_probs[i] for i in batch_indices])).to(device),
                'advantages': torch.FloatTensor(advantages[batch_indices]).to(device),
                'returns': torch.FloatTensor(returns[batch_indices]).to(device)
            }

    def get_top_k_per_path(self, num_paths, k=20):
        path_samples = {i: [] for i in range(num_paths)}

        for idx in range(len(self.states)):
            path_idx = self.path_indices[idx]
            similarity = self.similarities[idx]
            state = self.states[idx]
            path_samples[path_idx].append((idx, similarity, state))

        top_k_results = {}
        for path_idx in range(num_paths):
            samples = path_samples[path_idx]
            if len(samples) == 0:
                top_k_results[path_idx] = []
                continue

            samples.sort(key=lambda x: x[1], reverse=True)
            top_k = samples[:k]

            top_k_results[path_idx] = []
            for sample in top_k:
                normalized_state = sample[2]
                original_state = denormalize_state(normalized_state)
                original_state_int = np.round(original_state).astype(int)

                top_k_results[path_idx].append({
                    'state': original_state_int,
                    'similarity': sample[1],
                    'triggered': execute_Tr(*original_state_int)
                })

        return top_k_results

    def clear(self):
        self.states.clear()
        self.actions.clear()
        self.rewards.clear()
        self.values.clear()
        self.log_probs.clear()
        self.dones.clear()
        self.path_indices.clear()
        self.similarities.clear()

    def __len__(self):
        return len(self.states)

# === PPO ===
class PPOAgent:
    def __init__(self, state_dim=3, action_dim=3):
        self.state_dim = state_dim
        self.action_dim = action_dim

        self.actor = PPOActor(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.actor_optimizer = optim.Adam(self.actor.parameters(), lr=EXPERIMENT_CONFIG['ACTOR_LR'])

        self.critic = PPOCritic(state_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_optimizer = optim.Adam(self.critic.parameters(), lr=EXPERIMENT_CONFIG['CRITIC_LR'])

        self.buffer = PPOBuffer()
        self.update_count = 0

    def get_action(self, state):
            min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
            max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
            normalized_state = (state - (min_vals + max_vals) / 2) / ((max_vals - min_vals) / 2)
            state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)

            with torch.no_grad():
                action, log_prob = self.actor.get_action_and_log_prob(state_tensor)
                value = self.critic(state_tensor)

            action = action.cpu().numpy()[0]
            log_prob = log_prob.cpu().item()
            value = value.cpu().item()

            return action, log_prob, value
    def store_experience(self, state, action, reward, value, log_prob, done, path_idx, similarity):
            min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
            max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
            normalized_state = (state - (min_vals + max_vals) / 2) / ((max_vals - min_vals) / 2)
            self.buffer.store(normalized_state, action, reward, value, log_prob, done, path_idx, similarity)

    def update(self):
        if len(self.buffer) == 0:
            return

        advantages, returns = self.buffer.compute_advantages()

        for epoch in range(EXPERIMENT_CONFIG['PPO_EPOCHS']):
            for batch in self.buffer.get_batch_iterator(advantages, returns):
                new_log_probs = self.actor.get_log_prob(batch['states'], batch['actions'])
                ratio = torch.exp(new_log_probs - batch['old_log_probs'])

                surr1 = ratio * batch['advantages']
                surr2 = torch.clamp(ratio, 1 - EXPERIMENT_CONFIG['CLIP_EPSILON'],
                                    1 + EXPERIMENT_CONFIG['CLIP_EPSILON']) * batch['advantages']
                actor_loss = -torch.min(surr1, surr2).mean()

                self.actor_optimizer.zero_grad()
                actor_loss.backward()
                torch.nn.utils.clip_grad_norm_(self.actor.parameters(), 0.5)
                self.actor_optimizer.step()

                new_values = self.critic(batch['states'])
                critic_loss = F.mse_loss(new_values, batch['returns'])

                self.critic_optimizer.zero_grad()
                critic_loss.backward()
                torch.nn.utils.clip_grad_norm_(self.critic.parameters(), 0.5)
                self.critic_optimizer.step()

        self.update_count += 1
        self.buffer.clear()

        if self.update_count % 2 == 0:
            print(f"  -> PPOcompleted (Run {self.update_count})")

# === Metric ===
def calculate_run_performance(run_idx, ppo_results, training_time, total_steps, update_count, agent):
    """ runMetric"""
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    # 1. (Total Reward)
    total_reward = 0
    # 2. (Average Reward)
    average_reward = 0
    # 5. (Convergence)
    convergence = 0
    # 12. (Environment Adaptability)
    environment_adaptability = 0
    # 13. (Generalization Ability)
    generalization_ability = 0
    # 15. (Computational Efficiency)
    computational_efficiency = 0
    # 16. (Policy Update Frequency)
    policy_update_frequency = 0

    # Similarity
    all_similarities = []

    # Metric
    total_samples = 0
    all_rewards = []

    for path_idx in range(num_paths):
        samples = ppo_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']

            total_reward += reward
            all_rewards.append(reward)
            all_similarities.append(similarity)
            total_samples += 1

    # 1. 
    total_reward = total_reward

    # 2. 
    if total_samples > 0:
        average_reward = total_reward / total_samples

    # 5. (Average Similarity)
    if all_similarities:
        convergence = np.mean(all_similarities)

    # 12. (Similarity)
    if len(all_similarities) > 1:
        environment_adaptability = 1 / (np.std(all_similarities) + 1e-8)

    # 13. (Average Similarity)
    generalization_ability = convergence

    # 15. (/ seconds)
    if training_time > 0:
        computational_efficiency = total_steps / training_time

    # 16. 
    if training_time > 0:
        policy_update_frequency = update_count / training_time

    # Similarity
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0

    return {
        'Run': run_idx + 1,

        # Metric
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),

        # Similarity
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }

# === Excel ===
def export_to_excel(all_ppo_results, all_performance_data, target_paths, output_path="PPO_20 run.xlsx"):
    """20 runPPOExcel"""
    print("\nExcel...")

    # 
    all_ppo_summary_data = []
    all_ppo_detailed_data = []

    #  run
    for run_idx, (ppo_results, performance_data) in enumerate(zip(all_ppo_results, all_performance_data)):
        # ===== Sheet1: PPOPath  =====
        ppo_summary_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = ppo_results[path_idx]

            if len(samples) == 0:
                ppo_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue

            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = '' if perfect_count > 0 else ''

            ppo_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })

        all_ppo_summary_data.extend(ppo_summary_data)

        # ===== Sheet2: PPODetailed Sample Data =====
        ppo_detailed_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = ppo_results[path_idx]

            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']

                ppo_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })

        all_ppo_detailed_data.extend(ppo_detailed_data)

    # Excel
    ppo_summary_df = pd.DataFrame(all_ppo_summary_data)
    ppo_detailed_df = pd.DataFrame(all_ppo_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet1: PPOPath 
        ppo_summary_df.to_excel(writer, sheet_name='PPOPath ', index=False)

        # Sheet2: PPODetailed Sample Data
        ppo_detailed_df.to_excel(writer, sheet_name='PPODetailed Sample Data', index=False)

        # Sheet3: Metric - 
        selected_columns = [
            'Run',
            'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df_selected = performance_df[selected_columns]
        performance_df_selected.to_excel(writer, sheet_name='Metric', index=False)

        # 
        workbook = writer.book

        # 
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        perfect_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # 

        # === Sheet1 ===
        ws1 = writer.sheets['PPOPath ']
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 
        for row_idx in range(2, ws1.max_row + 1):
            if ws1.cell(row_idx, 9).value == '':  # Run 9""
                for col_idx in range(1, ws1.max_column + 1):
                    ws1.cell(row_idx, col_idx).fill = perfect_fill

        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 15
        ws1.column_dimensions['F'].width = 15
        ws1.column_dimensions['G'].width = 15
        ws1.column_dimensions['H'].width = 15
        ws1.column_dimensions['I'].width = 15
        ws1.column_dimensions['J'].width = 50

        # === Sheet2 ===
        ws2 = writer.sheets['PPODetailed Sample Data']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 15
        ws2.column_dimensions['I'].width = 40
        ws2.column_dimensions['J'].width = 40
        ws2.column_dimensions['K'].width = 15
        ws2.column_dimensions['L'].width = 15

        # === Sheet3 ===
        ws3 = writer.sheets['Metric']
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        for col in columns:
            ws3.column_dimensions[col].width = 18

    print(f"Excel: {output_path}")
    print(f"  - Sheet1: PPOPath  ({len(all_ppo_summary_data)})")
    print(f"  - Sheet2: PPODetailed Sample Data ({len(all_ppo_detailed_data)})")
    print(f"  - Sheet3: Metric ({len(all_performance_data)})")

# === PPO ===
def train_ppo_workflow():
    print("=" * 80)
    print("PPO")
    print("Similarity:  / target paths")
    print("=" * 80)

    agent = PPOAgent()
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n: Path {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']}")
    path_samples = {}
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']

    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            # 分别生成 X, Y, Z 的随机整数
            state = np.array([
                np.random.randint(min_vals[0], max_vals[0] + 1),
                np.random.randint(min_vals[1], max_vals[1] + 1),
                np.random.randint(min_vals[2], max_vals[2] + 1)
            ], dtype=np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  Path  {path_idx + 1}/{num_paths}:  {len(samples)} ")

    # 
    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size

    print(f"\n: {batch_size},{EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(f": {num_batches} /Path  x {num_paths} Path  = {num_batches * num_paths} ")
    print("-" * 80)

    global_buffer = PPOBuffer()

    for batch_idx in range(num_batches):
        print(f"\n {batch_idx + 1}/{num_batches}")

        for path_idx in range(num_paths):
            target_path = target_paths[path_idx]
            batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]

            batch_rewards = []
            batch_similarities = []

            for sample_idx, initial_state in enumerate(batch_samples):
                state = initial_state.copy()
                episode_reward = 0
                final_similarity = 0

                for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                    action, log_prob, value = agent.get_action(state)

                    next_state = state + action
                    next_state = clip_state(next_state)

                    triggered = execute_Tr(*next_state)
                    reward = unified_reward_function(triggered, target_path)
                    similarity = coverage_similarity(triggered, target_path)

                    done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)

                    agent.store_experience(state, action, reward, value, log_prob, done, path_idx, similarity)
                    normalized_state = (state - (min_vals + max_vals) / 2) / ((max_vals - min_vals) / 2)
                    global_buffer.store(
                        normalized_state,
                        action, reward, value, log_prob, done, path_idx, similarity
                    )

                    state = next_state
                    episode_reward += reward
                    final_similarity = similarity
                    total_steps += 1

                batch_rewards.append(episode_reward)
                batch_similarities.append(final_similarity)

            avg_reward = np.mean(batch_rewards)
            avg_similarity = np.mean(batch_similarities)
            print(f"  Path {path_idx + 1}: ={avg_reward:.2f}, Average Similarity={avg_similarity:.4f}")

        print(f"\n  PPO...")
        agent.update()
        print(f"  : {len(global_buffer)}")

    training_time = time.time() - start_time

    print("\n" + "=" * 80)
    print(f"PPOcompleted! Total elapsed time: {training_time:.2f} seconds, : {total_steps}")
    print(f": {len(global_buffer)}")
    print(f"PPO: {agent.update_count}")
    print("=" * 80)

    # Top-K
    print(f"\nPath SimilarityMaximum{EXPERIMENT_CONFIG['TOP_K_SAMPLES']}...")
    ppo_top_k_results = global_buffer.get_top_k_per_path(num_paths, EXPERIMENT_CONFIG['TOP_K_SAMPLES'])

    return agent, ppo_top_k_results, training_time, total_steps, agent.update_count

# ===  ===
def main():
    print("\n" + "=" * 80)
    print("PPO - 20 run")
    print("Metric")
    print("=" * 80)

    all_ppo_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    # 20
    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'=' * 80}")
        print(f"Start run  {run_idx + 1}/{EXPERIMENT_CONFIG['NUM_RUNS']}  run")
        print(f"{'=' * 80}")

        # PPO
        ppo_agent, ppo_results, training_time, total_steps, update_count = train_ppo_workflow()

        # Metric
        performance_data = calculate_run_performance(
            run_idx, ppo_results, training_time, total_steps, update_count, ppo_agent
        )

        # 
        all_ppo_results.append(ppo_results)
        all_performance_data.append(performance_data)

        print(f"\nRun  {run_idx + 1}  runcompleted!")
        print(f"  Total Reward: {performance_data['Total Reward']}")
        print(f"  Average Reward: {performance_data['Average Reward']}")
        print(f"  Average Similarity: {performance_data['Average Similarity']}")

    # Excel(20 run)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"PPO_20 run_{timestamp}.xlsx"
    export_to_excel(all_ppo_results, all_performance_data, target_paths, output_path)

    # 
    print("\n" + "=" * 80)
    print("20 run")
    print("=" * 80)

    # Metric
    total_rewards = [p['Total Reward'] for p in all_performance_data]
    average_rewards = [p['Average Reward'] for p in all_performance_data]
    convergences = [p['Convergence'] for p in all_performance_data]
    environment_adaptabilities = [p['Environment Adaptability'] for p in all_performance_data]
    generalization_abilities = [p['Generalization Ability'] for p in all_performance_data]
    computational_efficiencies = [p['Computational Efficiency'] for p in all_performance_data]
    policy_update_frequencies = [p['Policy Update Frequency'] for p in all_performance_data]
    avg_similarities = [p['Average Similarity'] for p in all_performance_data]

    print(f"\nTotal Reward Statistics:")
    print(f"  Mean: {np.mean(total_rewards):.2f}")
    print(f"  Standard deviation: {np.std(total_rewards):.2f}")

    print(f"\nAverage Reward Statistics:")
    print(f"  Mean: {np.mean(average_rewards):.4f}")
    print(f"  Standard deviation: {np.std(average_rewards):.4f}")

    print(f"\nConvergence Statistics:")
    print(f"  Mean: {np.mean(convergences):.4f}")
    print(f"  Standard deviation: {np.std(convergences):.4f}")

    print(f"\nEnvironment Adaptability Statistics:")
    print(f"  Mean: {np.mean(environment_adaptabilities):.4f}")
    print(f"  Standard deviation: {np.std(environment_adaptabilities):.4f}")

    print(f"\nGeneralization Ability Statistics:")
    print(f"  Mean: {np.mean(generalization_abilities):.4f}")
    print(f"  Standard deviation: {np.std(generalization_abilities):.4f}")

    print(f"\nComputational Efficiency Statistics:")
    print(f"  Mean: {np.mean(computational_efficiencies):.2f}")
    print(f"  Standard deviation: {np.std(computational_efficiencies):.2f}")

    print(f"\nPolicy Update Frequency Statistics:")
    print(f"  Mean: {np.mean(policy_update_frequencies):.4f}")
    print(f"  Standard deviation: {np.std(policy_update_frequencies):.4f}")

    print("\n" + "=" * 80)
    print(f" {EXPERIMENT_CONFIG['NUM_RUNS']} completed!")
    print("=" * 80)

if __name__ == "__main__":
    main()