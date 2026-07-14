import os
import sys
import random
import time
from collections import deque
from datetime import datetime
from typing import Any, Dict, List, Optional, Set, Tuple

import numpy as np
import torch
import torch.nn as nn
import torch.optim as optim
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ========================================
# ========== experiment configuration parameters ==========
# ========================================
NUM_RUNS = 20

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === three-dimensional range settings ===
# Keep the current DQN state range used by the second script. To use a 0-500 range, modify this section only.
LIGHT_MIN = 1
LIGHT_MAX = 6
MOISTURE_MIN = 1
MOISTURE_MAX = 6
TEMP_MIN = 1
TEMP_MAX = 6
BOUNDS = {
    "light": (LIGHT_MIN, LIGHT_MAX),
    "moisture": (MOISTURE_MIN, MOISTURE_MAX),
    "temp": (TEMP_MIN, TEMP_MAX),
}

# === standard PSO parameters ===
PSO_W = 0.7
PSO_C1 = 1.5
PSO_C2 = 1.5
PSO_VMAX_RATIO = 0.2


def normalize_state(state):
    """Normalize the state to the [0, 1] range."""
    return np.array([
        (state[0] - BOUNDS["light"][0]) / (BOUNDS["light"][1] - BOUNDS["light"][0]),
        (state[1] - BOUNDS["moisture"][0]) / (BOUNDS["moisture"][1] - BOUNDS["moisture"][0]),
        (state[2] - BOUNDS["temp"][0]) / (BOUNDS["temp"][1] - BOUNDS["temp"][0]),
    ], dtype=np.float32)


def denormalize_state(normalized_state):
    """Restore the normalized state to the original range."""
    return np.array([
        normalized_state[0] * (BOUNDS["light"][1] - BOUNDS["light"][0]) + BOUNDS["light"][0],
        normalized_state[1] * (BOUNDS["moisture"][1] - BOUNDS["moisture"][0]) + BOUNDS["moisture"][0],
        normalized_state[2] * (BOUNDS["temp"][1] - BOUNDS["temp"][0]) + BOUNDS["temp"][0],
    ], dtype=float)


def generate_random_state():
    """Generate a random state within each dimension range."""
    light = np.random.randint(BOUNDS["light"][0], BOUNDS["light"][1] + 1)
    moisture = np.random.randint(BOUNDS["moisture"][0], BOUNDS["moisture"][1] + 1)
    temp = np.random.randint(BOUNDS["temp"][0], BOUNDS["temp"][1] + 1)
    return np.array([light, moisture, temp], dtype=float)


def clip_state(state):
    """Clip the state to each dimension boundary."""
    return np.array([
        np.clip(state[0], BOUNDS["light"][0], BOUNDS["light"][1]),
        np.clip(state[1], BOUNDS["moisture"][0], BOUNDS["moisture"][1]),
        np.clip(state[2], BOUNDS["temp"][0], BOUNDS["temp"][1]),
    ], dtype=float)


def is_state_valid(state):
    """Check whether the state is within all valid dimension ranges."""
    return (
        BOUNDS["light"][0] <= state[0] <= BOUNDS["light"][1]
        and BOUNDS["moisture"][0] <= state[1] <= BOUNDS["moisture"][1]
        and BOUNDS["temp"][0] <= state[2] <= BOUNDS["temp"][1]
    )


def execute_Tr(weather, time_period, z):
    """验证规则函数 - weather, time_period, z组合（删除覆盖率为0的分支后重新编号版本）"""

    triggered = set()

    # 将z映射到合适的范围以匹配条件逻辑
    x = z  # 直接使用z作为x
    y = (weather * time_period * 10 + z) % 100 + 1  # 基于输入参数计算y值

    # 1-7: 早高峰组合（time_period == 1）
    if time_period == 1:
        if (x < 60 and y > 75) != (x < 60 and y < 75):
            triggered.add(1)
        if (x > 60 and y > 70) != (x < 60 and y > 70):
            triggered.add(2)
        if (x < 50 and y < 40) != (x > 50 and y < 40):
            triggered.add(3)
        if (x > 78 and 45 < y < 70) != (x < 78 and 45 < y < 70):
            triggered.add(4)
        if (45 < x < 70 and y > 78) != (45 < x < 70 and y < 78):
            triggered.add(5)
        if (x < 55 and 50 < y < 75) != (x > 55 and 50 < y < 75):
            triggered.add(6)
        if (50 < x < 75 and y < 55) != (50 < x < 75 and y > 55):
            triggered.add(7)

    # 8-14: 晚高峰组合（time_period == 2）
    if time_period == 2:
        if (x < 60 and y > 75) != (x < 60 and y < 75):
            triggered.add(8)
        if (x > 60 and y > 70) != (x < 60 and y > 70):
            triggered.add(9)
        if (x < 55 and y < 45) != (x > 55 and y < 45):
            triggered.add(10)
        if (45 < x < 70 and y > 78) != (45 < x < 70 and y < 78):
            triggered.add(11)
        if (x > 78 and 45 < y < 70) != (x < 78 and 45 < y < 70):
            triggered.add(12)
        if (55 < x < 75 and y < 50) != (55 < x < 75 and y > 50):
            triggered.add(13)
        if (x < 50 and 55 < y < 75) != (x > 50 and 55 < y < 75):
            triggered.add(14)

    # 15-19: 午餐时间组合（time_period == 3）
    if time_period == 3:
        if (x > 60 and 40 < y < 65) != (x < 60 and 40 < y < 65):
            triggered.add(15)
        if (40 < x < 65 and y > 60) != (40 < x < 65 and y < 60):
            triggered.add(16)
        if (45 < x < 70 and 45 < y < 60) != (45 < x < 70 and y > 60):
            triggered.add(17)  # Fixed: 45 > y > 60 was invalid
        if (x < 50 and y < 40) != (x > 50 and y < 40):
            triggered.add(18)
        if (x > 65 and y < 45) != (x < 65 and y < 45):
            triggered.add(19)

    # 20-25: 夜间组合（time_period == 4）
    if time_period == 4:
        if (x < 45 and y < 35) != (x > 45 and y < 35):
            triggered.add(20)
        if (x > 60 and y < 40) != (x < 60 and y < 40):
            triggered.add(21)
        if (x < 50 and y > 70) != (x > 50 and y > 70):
            triggered.add(22)
        if (45 < x < 70 and 45 < y < 60) != (x > 70 and 45 < y < 60):
            triggered.add(23)  # Fixed: 45 > x > 70 was invalid
        if (x < 35 and y < 25) != (x > 35 and y < 25):
            triggered.add(24)
        if (40 < x < 65 and y < 45) != (40 < x < 65 and y > 45):
            triggered.add(25)

    # 26-28: 周末组合（time_period == 5）
    if time_period == 5:
        if (x < 60 and y < 50) != (x > 60 and y < 50):
            triggered.add(26)
        if (x > 65 and y > 75) != (x < 65 and y > 75):
            triggered.add(27)
        if (x > 60 and y < 45) != (x < 60 and y < 45):
            triggered.add(28)

    # 29-33: 假日组合（time_period == 6）
    if time_period == 6:
        if (40 < x < 70 and 40 < y < 60) != (x > 70 and 40 < y < 60):
            triggered.add(29)  # Fixed: 40 > x > 70 was invalid
        if (x < 55 and y < 45) != (x > 55 and y < 45):
            triggered.add(30)
        if (x > 60 and y < 50) != (x < 60 and y < 50):
            triggered.add(31)
        if (x < 60 and y > 70) != (x > 60 and y > 70):
            triggered.add(32)
        if (x > 65 and y > 75) != (x < 65 and y > 75):
            triggered.add(33)

    # 34-68: 天气相关扩展规则
    if weather == 1:  # 晴天
        if (time_period in [1, 2] and x > 70) != (time_period in [1, 2] and x < 70):
            triggered.add(34)
        if (time_period in [1, 2] and y > 70) != (time_period in [1, 2] and y < 70):
            triggered.add(35)
        if (time_period in [3, 4] and x < 50) != (time_period in [3, 4] and x > 50):
            triggered.add(36)
        if (time_period in [3, 4] and y < 50) != (time_period in [3, 4] and y > 50):
            triggered.add(37)
        if (time_period in [5, 6] and 40 < x < 80) != (time_period in [5, 6] and x > 80):
            triggered.add(38)  # Fixed: 40 > x > 80 was invalid
        if (time_period in [5, 6] and 40 < y < 80) != (time_period in [5, 6] and y > 80):
            triggered.add(39)  # Fixed: 40 > y > 80 was invalid

    if weather == 2:  # 雨天
        if (time_period in [1, 2] and x > 75) != (time_period in [1, 2] and x < 75):
            triggered.add(40)
        if (time_period in [1, 2] and y < 60) != (time_period in [1, 2] and y > 60):
            triggered.add(41)
        if (time_period in [3, 4] and x < 45) != (time_period in [3, 4] and x > 45):
            triggered.add(42)
        if (time_period in [3, 4] and y > 65) != (time_period in [3, 4] and y < 65):
            triggered.add(43)
        if (time_period in [5, 6] and 35 < x < 75) != (time_period in [5, 6] and x > 75):
            triggered.add(44)  # Fixed: 35 > x > 75 was invalid
        if (time_period in [5, 6] and 35 < y < 75) != (time_period in [5, 6] and y > 75):
            triggered.add(45)  # Fixed: 35 > y > 75 was invalid

    if weather == 3:  # 雾天
        if (time_period in [1, 2] and x > 60) != (time_period in [1, 2] and x < 60):
            triggered.add(46)
        if (time_period in [1, 2] and y > 65) != (time_period in [1, 2] and y < 65):
            triggered.add(47)
        if (time_period in [3, 4] and x < 55) != (time_period in [3, 4] and x > 55):
            triggered.add(48)
        if (time_period in [3, 4] and y < 55) != (time_period in [3, 4] and y > 55):
            triggered.add(49)
        if (time_period in [5, 6] and 30 < x < 70) != (time_period in [5, 6] and x > 70):
            triggered.add(50)  # Fixed: 30 > x > 70 was invalid
        if (time_period in [5, 6] and 30 < y < 70) != (time_period in [5, 6] and y > 70):
            triggered.add(51)  # Fixed: 30 > y > 70 was invalid

    if weather == 4:  # 雪天
        if (time_period in [1, 2] and x > 65) != (time_period in [1, 2] and x < 65):
            triggered.add(52)
        if (time_period in [1, 2] and y < 55) != (time_period in [1, 2] and y > 55):
            triggered.add(53)
        if (time_period in [3, 4] and x < 40) != (time_period in [3, 4] and x > 40):
            triggered.add(54)
        if (time_period in [3, 4] and y > 60) != (time_period in [3, 4] and y < 60):
            triggered.add(55)
        if (time_period in [5, 6] and 25 < x < 65) != (time_period in [5, 6] and x > 65):
            triggered.add(56)  # Fixed: 25 > x > 65 was invalid
        if (time_period in [5, 6] and 25 < y < 65) != (time_period in [5, 6] and y > 65):
            triggered.add(57)  # Fixed: 25 > y > 65 was invalid

    if weather == 5:  # 风天
        if (time_period in [1, 2] and x > 70) != (time_period in [1, 2] and x < 70):
            triggered.add(58)
        if (time_period in [1, 2] and y > 60) != (time_period in [1, 2] and y < 60):
            triggered.add(59)
        if (time_period in [3, 4] and x < 35) != (time_period in [3, 4] and x > 35):
            triggered.add(60)
        if (time_period in [3, 4] and y < 40) != (time_period in [3, 4] and y > 40):
            triggered.add(61)
        if (time_period in [5, 6] and 20 < x < 60) != (time_period in [5, 6] and x > 60):
            triggered.add(62)  # Fixed: 20 > x > 60 was invalid
        if (time_period in [5, 6] and 20 < y < 60) != (time_period in [5, 6] and y > 60):
            triggered.add(63)  # Fixed: 20 > y > 60 was invalid

    if weather == 6:  # 暴雨
        if (time_period in [1, 2] and x > 55) != (time_period in [1, 2] and x < 55):
            triggered.add(64)
        if (time_period in [1, 2] and y > 55) != (time_period in [1, 2] and y < 55):
            triggered.add(65)
        if (time_period in [3, 4] and x < 45) != (time_period in [3, 4] and x > 45):
            triggered.add(66)
        if (time_period in [3, 4] and y < 45) != (time_period in [3, 4] and y > 45):
            triggered.add(67)
        if (time_period in [5, 6] and 15 < x < 55) != (time_period in [5, 6] and x > 55):
            triggered.add(68)  # Fixed: 15 > x > 55 was invalid

    # 69-78: 复合条件（多参数组合）
    if weather + time_period > 6:
        if (x > 50 and y > 50) != (x < 50 and y > 50):
            triggered.add(69)
        if (x < 50 and y < 50) != (x > 50 and y < 50):
            triggered.add(70)
        if (x > y) != (x < y):
            triggered.add(71)
        if (x < y) != (x > y):
            triggered.add(72)
        if (abs(x - y) < 20) != (abs(x - y) > 20):
            triggered.add(73)

    if weather + time_period <= 6:
        if (x > 60 or y > 60) != (x < 60 or y > 60):
            triggered.add(74)
        if (x < 40 or y < 40) != (x > 40 or y < 40):
            triggered.add(75)
        if (x + y > 100) != (x + y < 100):
            triggered.add(76)
        if (x + y < 80) != (x + y > 80):
            triggered.add(77)
        if (abs(x - y) > 30) != (abs(x - y) < 30):
            triggered.add(78)

    # 79-88: 数值关系条件
    if weather % 2 == time_period % 2:  # 同奇偶性
        if (x % 10 < 5) != (x % 10 > 5):
            triggered.add(79)
        if (y % 10 >= 5) != (y % 10 < 5):
            triggered.add(80)
        if ((x + y) % 3 == 0) != ((x + y) % 3 == 1):
            triggered.add(81)
        if ((x * y) % 7 == 0) != ((x * y) % 7 == 1):
            triggered.add(82)
        if (x // 10 == y // 10) != (x // 10 != y // 10):
            triggered.add(83)

    if weather % 2 != time_period % 2:  # 不同奇偶性
        if (x > 75 or y > 75) != (x < 75 or y > 75):
            triggered.add(84)
        if (x < 25 or y < 25) != (x > 25 or y < 25):
            triggered.add(85)
        if (max(x, y) - min(x, y) > 40) != (max(x, y) - min(x, y) < 40):
            triggered.add(86)
        if ((x + y) // 2 > 50) != ((x + y) // 2 < 50):
            triggered.add(87)
        if (weather * time_period > 15) != (weather * time_period < 15):
            triggered.add(88)

    # 89-95: 高级组合条件（奇数天气）
    if weather in [1, 3, 5]:  # 奇数天气
        if (time_period in [1, 3, 5] and x > 40) != (time_period in [1, 3, 5] and x < 40):
            triggered.add(89)
        if (time_period in [2, 4, 6] and y > 40) != (time_period in [2, 4, 6] and y < 40):
            triggered.add(90)
        if (x % 20 < 10 and y % 20 < 10) != (x % 20 > 10 and y % 20 < 10):
            triggered.add(91)
        if (x + weather * 10 > 50) != (x + weather * 10 < 50):
            triggered.add(92)
        if (y + time_period * 10 > 50) != (y + time_period * 10 < 50):
            triggered.add(93)
        if (time_period in [1, 3, 5] and x < 60) != (time_period in [1, 3, 5] and x > 60):
            triggered.add(94)
        if (time_period in [2, 4, 6] and y < 60) != (time_period in [2, 4, 6] and y > 60):
            triggered.add(95)

    # 96-98: 偶数天气条件
    if weather in [2, 4, 6]:  # 偶数天气
        if ((x + y) % weather == 0) != ((x + y) % weather == 1):
            triggered.add(96)
        if (x * weather > 100) != (x * weather < 100):
            triggered.add(97)
        if (y * time_period > 100) != (y * time_period < 100):
            triggered.add(98)

    # 99-100: 最后的复杂条件
    if ((weather * time_period + z) % 7 == 0) != ((weather * time_period + z) % 7 == 1):
        triggered.add(99)
    if (max(weather, time_period) * min(x, y) > 150) != (max(weather, time_period) * min(x, y) < 150):
        triggered.add(100)

    return triggered

target_paths = [
    {15, 16, 48, 49, 74, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 91, 92, 93, 94, 99, 100},
    {16, 18, 19, 60, 61, 70, 71, 72, 73, 80, 81, 82, 83, 89, 91, 92, 93, 94, 99, 100},
    {1, 4, 6, 46, 47, 74, 75, 76, 77, 78, 80, 81, 82, 83, 89, 92, 93, 94, 99, 100},
    {30, 31, 50, 51, 70, 71, 72, 73, 84, 85, 86, 87, 88, 91, 92, 93, 95, 99, 100},
    {18, 19, 36, 37, 74, 76, 77, 78, 79, 80, 81, 82, 83, 89, 92, 93, 94, 99, 100},
    {20, 24, 25, 36, 37, 76, 77, 78, 84, 86, 87, 88, 90, 91, 92, 93, 95, 99, 100},
    {8, 12, 34, 35, 74, 75, 76, 77, 78, 84, 86, 87, 88, 90, 91, 92, 93, 95, 100},
    {8, 10, 58, 59, 70, 71, 72, 73, 84, 85, 86, 87, 88, 91, 92, 93, 95, 99, 100},
    {8, 14, 46, 47, 75, 76, 77, 78, 84, 85, 86, 87, 88, 90, 92, 93, 95, 99, 100},
    {1, 2, 6, 46, 47, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 92, 93, 94, 100},
    {39, 74, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 91, 92, 93, 94, 99, 100},
    {20, 21, 60, 61, 70, 71, 72, 73, 84, 85, 86, 87, 88, 90, 92, 93, 95, 99},
    {8, 9, 11, 13, 40, 41, 75, 76, 77, 78, 79, 80, 81, 83, 96, 97, 98, 100},
    {18, 19, 54, 55, 70, 71, 72, 73, 84, 86, 87, 88, 96, 97, 98, 99, 100},
    {27, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 91, 92, 93, 94, 99, 100},
    {25, 48, 49, 69, 71, 72, 73, 84, 85, 86, 87, 88, 90, 92, 93, 95, 100},
    {26, 28, 62, 70, 71, 72, 73, 80, 81, 82, 83, 89, 91, 92, 93, 94, 100},
    {32, 33, 68, 69, 71, 72, 73, 79, 80, 81, 82, 83, 96, 97, 98, 99, 100},
    {1, 52, 53, 74, 75, 76, 77, 78, 84, 85, 86, 87, 88, 97, 98, 99, 100},
    {8, 12, 14, 64, 65, 69, 71, 72, 73, 80, 81, 82, 83, 96, 97, 98, 100},
    {1, 3, 64, 65, 70, 71, 72, 73, 84, 86, 87, 88, 96, 97, 98, 99, 100},
    {22, 36, 37, 76, 77, 78, 85, 86, 87, 88, 90, 91, 93, 95, 100},
    {31, 45, 70, 71, 72, 73, 79, 80, 81, 83, 96, 97, 98, 99, 100},
    {22, 66, 67, 69, 71, 72, 73, 79, 80, 82, 83, 97, 98, 100},
    {44, 45, 69, 71, 72, 73, 79, 80, 83, 96, 97, 98, 99, 100},
    {57, 71, 72, 73, 79, 80, 83, 97, 98, 100},
    {15, 16, 17, 48, 49, 74, 75, 76, 77, 78, 79, 80, 82, 83, 89, 91, 92, 93, 94, 100},
    {1, 2, 5, 46, 47, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 91, 92, 93, 94, 100},
    {20, 21, 25, 42, 43, 74, 76, 77, 78, 79, 80, 81, 82, 83, 96, 97, 98, 99, 100},
    {2, 5, 7, 40, 41, 75, 76, 77, 78, 84, 85, 86, 87, 88, 96, 97, 98, 99, 100},
    {26, 28, 56, 57, 70, 71, 72, 73, 84, 85, 86, 87, 88, 96, 97, 98, 99, 100},
    {26, 28, 38, 74, 76, 77, 78, 80, 81, 82, 83, 89, 91, 92, 93, 94, 100},
    {30, 31, 62, 63, 70, 71, 72, 73, 84, 86, 87, 88, 90, 91, 92, 93, 95},
    {29, 62, 63, 71, 72, 73, 84, 85, 86, 87, 88, 90, 92, 93, 95, 100},
    {23, 25, 60, 61, 71, 72, 73, 84, 85, 86, 87, 88, 90, 92, 93, 95, 100}
]


def jaccard_similarity(set1: Set[int], set2: Set[int]) -> float:
    """If set1 covers target set2, the match score is treated as 1."""
    if set2.issubset(set1):
        return 1.0
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    return intersection / union if union != 0 else 0.0


def compute_reward(state, target_path, triggered):
    """Compute the DQN reward."""
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward


class SimpleReplayBuffer:
    def __init__(self, capacity=5000):
        self.buffer = deque(maxlen=capacity)

    def append(self, experience):
        self.buffer.append(experience)

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return random.sample(self.buffer, len(self.buffer))
        return random.sample(self.buffer, batch_size)

    def __len__(self):
        return len(self.buffer)


class SimpleDQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super().__init__()
        self.fc1 = nn.Linear(state_dim, 64)
        self.fc2 = nn.Linear(64, 32)
        self.fc3 = nn.Linear(32, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)


class SimpleDQNAgent:
    def __init__(self, state_dim, action_dim, learning_rate=0.001):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.epsilon = 0.3
        self.gamma = 0.99

        self.model = SimpleDQN(state_dim, action_dim).to(device)
        self.target_model = SimpleDQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())
        self.replay_buffer = SimpleReplayBuffer(capacity=5000)

    def decode_action(self, action_idx):
        """Decode the action."""
        light_deltas = [int((LIGHT_MAX - LIGHT_MIN) * p) for p in [0.05, 0.1, 0.2, 0.5, 0.7, -0.05, -0.1, -0.2, -0.5, -0.7]]
        moisture_deltas = [int((MOISTURE_MAX - MOISTURE_MIN) * p) for p in [0.05, 0.1, 0.2, 0.5, 0.7, -0.05, -0.1, -0.2, -0.5, -0.7]]
        temp_deltas = [int((TEMP_MAX - TEMP_MIN) * p) for p in [0.05, 0.1, 0.2, 0.5, 0.7, -0.05, -0.1, -0.2, -0.5, -0.7]]

        dim = action_idx // 10
        delta_idx = action_idx % 10

        if dim == 0:
            return (light_deltas[delta_idx], 0, 0)
        if dim == 1:
            return (0, moisture_deltas[delta_idx], 0)
        return (0, 0, temp_deltas[delta_idx])

    def store_transition(self, state, action, reward, next_state, done):
        self.replay_buffer.append((state, action, reward, next_state, done))

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return

        batch = self.replay_buffer.sample(batch_size)
        states, actions, rewards, next_states, dones = zip(*batch)

        states = torch.tensor(np.array(states), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array(next_states), dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))

        loss = nn.MSELoss()(current_q_values, target_q_values)

        self.optimizer.zero_grad()
        loss.backward()
        self.optimizer.step()

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())

    def get_best_samples(self, target_path, num_samples=20):
        """Screen DQN candidate samples from the replay buffer. This stage is counted separately in T_sample_screening."""
        if len(self.replay_buffer) == 0:
            return []

        samples_with_scores = []
        for state_norm, _, _, _, _ in self.replay_buffer.buffer:
            state_original = denormalize_state(state_norm)
            triggered = execute_Tr(state_original)
            reward = compute_reward(state_original, target_path, triggered)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_scores.append((state_original, reward, sim, triggered))

        samples_with_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_scores[:num_samples]


def train_dqn_for_path(path_idx, target_path, num_samples=200):
    """Train a standard DQN for a single path."""
    print(f"  Start training path {path_idx + 1} DQN model...")
    agent = SimpleDQNAgent(state_dim=3, action_dim=30)

    random_states = [normalize_state(generate_random_state()) for _ in range(num_samples)]

    STEPS_PER_SAMPLE = 3
    EPOCHS = 5
    BATCH_SIZE = 32

    step_count = 0
    for _ in range(EPOCHS):
        for state_norm in random_states:
            state_norm = tuple(state_norm)
            state_original = denormalize_state(state_norm)

            for step in range(STEPS_PER_SAMPLE):
                legal_actions = []
                for a in range(agent.action_dim):
                    dx, dy, dz = agent.decode_action(a)
                    next_state_candidate = (state_original[0] + dx, state_original[1] + dy, state_original[2] + dz)
                    if is_state_valid(next_state_candidate):
                        legal_actions.append(a)

                if not legal_actions:
                    break

                if random.random() < agent.epsilon:
                    action = random.choice(legal_actions)
                else:
                    state_tensor = torch.tensor(state_norm, dtype=torch.float32).unsqueeze(0).to(device)
                    with torch.no_grad():
                        q_values = agent.model(state_tensor)[0]
                    action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                dx, dy, dz = agent.decode_action(action)
                next_state_original = (state_original[0] + dx, state_original[1] + dy, state_original[2] + dz)
                next_state_norm = normalize_state(next_state_original)

                triggered = execute_Tr(next_state_original)
                reward = compute_reward(next_state_original, target_path, triggered)
                done = (step == STEPS_PER_SAMPLE - 1)

                agent.store_transition(state_norm, action, reward, next_state_norm, done)

                state_norm = next_state_norm
                state_original = next_state_original
                step_count += 1

                if step_count % 50 == 0 and len(agent.replay_buffer) >= BATCH_SIZE:
                    agent.train(BATCH_SIZE)

                if step_count % 100 == 0:
                    agent.update_target_model()

    print(f"  Path {path_idx + 1} training completed, replay buffer size: {len(agent.replay_buffer)}")
    return agent


class Particle:
    def __init__(self, initial_position=None):
        if initial_position is not None:
            self.position = clip_state(np.array(initial_position, dtype=float))
        else:
            self.position = np.array([
                np.random.uniform(BOUNDS["light"][0], BOUNDS["light"][1]),
                np.random.uniform(BOUNDS["moisture"][0], BOUNDS["moisture"][1]),
                np.random.uniform(BOUNDS["temp"][0], BOUNDS["temp"][1]),
            ], dtype=float)

        spans = np.array([
            BOUNDS["light"][1] - BOUNDS["light"][0],
            BOUNDS["moisture"][1] - BOUNDS["moisture"][0],
            BOUNDS["temp"][1] - BOUNDS["temp"][0],
        ], dtype=float)
        self.velocity = np.random.uniform(-0.1 * spans, 0.1 * spans)
        self.best_position = self.position.copy()
        self.best_fitness = -1.0
        self.fitness = -1.0


class BasicPSO:
    """
    Standard PSO: contains only standard velocity and position updates.
    Mutation coefficients, opposite particles, local-best reset, and other improved PSO mechanisms have been removed.
    """

    def __init__(self, target_path, swarm_size=20, dqn_samples=None):
        self.target_path = target_path
        self.swarm_size = swarm_size
        self.particles = []
        self.global_best_position = None
        self.global_best_fitness = -1.0

        if dqn_samples is not None and len(dqn_samples) > 0:
            num_direct = min(len(dqn_samples), swarm_size)
            for i in range(num_direct):
                state_tuple, _, _, _ = dqn_samples[i]
                self.particles.append(Particle(initial_position=state_tuple))

            while len(self.particles) < swarm_size:
                base_idx = len(self.particles) % len(dqn_samples)
                state_tuple, _, _, _ = dqn_samples[base_idx]
                perturbed = np.array(state_tuple, dtype=float) + np.random.randint(-10, 11, size=3)
                self.particles.append(Particle(initial_position=clip_state(perturbed)))
        else:
            self.particles = [Particle() for _ in range(swarm_size)]

        for particle in self.particles:
            particle.fitness = self.fitness_function(particle.position)
            particle.best_fitness = particle.fitness
            particle.best_position = particle.position.copy()

            if particle.fitness > self.global_best_fitness:
                self.global_best_fitness = particle.fitness
                self.global_best_position = particle.position.copy()

    def fitness_function(self, position):
        try:
            triggered = execute_Tr(position)
            return jaccard_similarity(triggered, self.target_path)
        except Exception:
            return 0.0

    def update(self):
        spans = np.array([
            BOUNDS["light"][1] - BOUNDS["light"][0],
            BOUNDS["moisture"][1] - BOUNDS["moisture"][0],
            BOUNDS["temp"][1] - BOUNDS["temp"][0],
        ], dtype=float)
        max_velocity = PSO_VMAX_RATIO * spans

        for particle in self.particles:
            r1 = np.random.random(3)
            r2 = np.random.random(3)

            particle.velocity = (
                PSO_W * particle.velocity
                + PSO_C1 * r1 * (particle.best_position - particle.position)
                + PSO_C2 * r2 * (self.global_best_position - particle.position)
            )
            particle.velocity = np.clip(particle.velocity, -max_velocity, max_velocity)

            particle.position = clip_state(particle.position + particle.velocity)
            particle.fitness = self.fitness_function(particle.position)

            if particle.fitness > particle.best_fitness:
                particle.best_fitness = particle.fitness
                particle.best_position = particle.position.copy()

            if particle.fitness > self.global_best_fitness:
                self.global_best_fitness = particle.fitness
                self.global_best_position = particle.position.copy()


def _set_header(ws, headers, widths, header_fill, header_font, border, center_align):
    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_row(ws, row_idx, row_data, border, alignment, fill=None):
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.border = border
        cell.alignment = alignment
        if fill is not None:
            cell.fill = fill


def export_time_metrics_to_excel(all_run_results, experiment_total_time=None, filename=None):
    """Export only timing metrics and per-path iteration counts."""
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"DQN_BasicPSO_Time_Metrics_{timestamp}.xlsx"

    output_dir = os.path.join(os.getcwd(), "results")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    num_runs = len(all_run_results)
    num_paths = len(target_paths)

    # Worksheet 1: runtime summary
    ws1 = wb.active
    ws1.title = "Runtime Summary"
    ws1.sheet_view.showGridLines = False
    headers1 = [
        "Run", "Number of Paths", "DQNTraining Total Time(s)", "Sample Screening Total Time(s)", "DQNDirect Check Total Time(s)",
        "PSOInitialization Total Time(s)", "PSOSearch Total Time(s)", "Total Algorithm Time(s)", "Total Iterations", "Average Iterations per Path"
    ]
    widths1 = [12, 10, 18, 18, 22, 18, 18, 16, 14, 18]
    _set_header(ws1, headers1, widths1, header_fill, header_font, border, center_align)

    for run_idx, run_result in enumerate(all_run_results, 1):
        path_results = run_result["paths"]
        total_dqn_train = sum(p["T_DQN_train"] for p in path_results)
        total_screen = sum(p["T_sample_screening"] for p in path_results)
        total_direct_check = sum(p["T_direct_check"] for p in path_results)
        total_pso_init = sum(p["T_PSO_init"] for p in path_results)
        total_pso_search = sum(p["T_PSO_search"] for p in path_results)
        total_iterations = sum(p["iterations"] for p in path_results)

        row_data = [
            f" {run_idx}",
            num_paths,
            round(total_dqn_train, 6),
            round(total_screen, 6),
            round(total_direct_check, 6),
            round(total_pso_init, 6),
            round(total_pso_search, 6),
            round(run_result["T_run_algorithm_total"], 6),
            int(total_iterations),
            round(total_iterations / num_paths, 2),
        ]
        _write_row(ws1, run_idx + 1, row_data, border, center_align, alternate_fill if run_idx % 2 == 0 else None)

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:J{num_runs + 1}"

    # Worksheet 2: path timing and iteration details
    ws2 = wb.create_sheet(title="Path Timing and Iteration Details")
    ws2.sheet_view.showGridLines = False
    headers2 = [
        "Run", "Path ID", "DQN training(s)", "Sample Screening Time(s)", "DQNDirect Check Time(s)",
        "PSOInitialization Time(s)", "PSOSearch Time(s)", "Path Total Algorithm Time(s)", "Iterations"
    ]
    widths2 = [12, 12, 18, 18, 22, 18, 18, 18, 12]
    _set_header(ws2, headers2, widths2, header_fill, header_font, border, center_align)

    row_idx = 2
    for run_idx, run_result in enumerate(all_run_results, 1):
        for path_result in run_result["paths"]:
            row_data = [
                f" {run_idx}",
                f"Path  {path_result['path_idx'] + 1}",
                round(path_result["T_DQN_train"], 6),
                round(path_result["T_sample_screening"], 6),
                round(path_result["T_direct_check"], 6),
                round(path_result["T_PSO_init"], 6),
                round(path_result["T_PSO_search"], 6),
                round(path_result["T_path_algorithm_total"], 6),
                int(path_result["iterations"]),
            ]
            _write_row(ws2, row_idx, row_data, border, center_align, alternate_fill if row_idx % 2 == 0 else None)
            row_idx += 1

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:I{row_idx - 1}"

    # Worksheet 3: path aggregate statistics
    ws3 = wb.create_sheet(title="Path Aggregate Statistics")
    ws3.sheet_view.showGridLines = False
    headers3 = [
        "Path ID", "Average DQN Training Time(s)", "Average Sample Screening Time(s)", "Average PSO Initialization Time(s)",
        "Average PSO Search Time(s)", "Average Path Total Algorithm Time(s)", "Average Iterations", "Minimum Iterations", "Maximum Iterations"
    ]
    widths3 = [12, 22, 22, 22, 22, 22, 16, 14, 14]
    _set_header(ws3, headers3, widths3, header_fill, header_font, border, center_align)

    for path_idx in range(num_paths):
        records = [run_result["paths"][path_idx] for run_result in all_run_results]
        iterations = [r["iterations"] for r in records]
        row_data = [
            f"Path  {path_idx + 1}",
            round(float(np.mean([r["T_DQN_train"] for r in records])), 6),
            round(float(np.mean([r["T_sample_screening"] for r in records])), 6),
            round(float(np.mean([r["T_PSO_init"] for r in records])), 6),
            round(float(np.mean([r["T_PSO_search"] for r in records])), 6),
            round(float(np.mean([r["T_path_algorithm_total"] for r in records])), 6),
            round(float(np.mean(iterations)), 2),
            int(np.min(iterations)),
            int(np.max(iterations)),
        ]
        _write_row(ws3, path_idx + 2, row_data, border, center_align, alternate_fill if (path_idx + 1) % 2 == 0 else None)

    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:I{num_paths + 1}"

    # Worksheet 4: experiment overview
    ws4 = wb.create_sheet(title="Experiment Overview")
    ws4.sheet_view.showGridLines = False
    headers4 = ["Metric", "Value"]
    widths4 = [32, 24]
    _set_header(ws4, headers4, widths4, header_fill, header_font, border, center_align)

    all_path_records = [p for run_result in all_run_results for p in run_result["paths"]]
    all_iterations = [p["iterations"] for p in all_path_records]
    overview_rows = [
        ["Experiment Runs", num_runs],
        ["Number of Target Paths", num_paths],
        ["Total Experiment Wall-clock Time(s)", round(experiment_total_time, 6) if experiment_total_time is not None else ""],
        ["Average Total Algorithm Time per Run(s)", round(float(np.mean([r["T_run_algorithm_total"] for r in all_run_results])), 6)],
        ["Average Total Algorithm Time per Path(s)", round(float(np.mean([p["T_path_algorithm_total"] for p in all_path_records])), 6)],
        ["Average DQN Training Time per Path(s)", round(float(np.mean([p["T_DQN_train"] for p in all_path_records])), 6)],
        ["Average Sample Screening Time per Path(s)", round(float(np.mean([p["T_sample_screening"] for p in all_path_records])), 6)],
        ["Average PSO Search Time per Path(s)", round(float(np.mean([p["T_PSO_search"] for p in all_path_records])), 6)],
        ["Average Iterations per Path", round(float(np.mean(all_iterations)), 2)],
        ["Maximum Iterations per Path", int(np.max(all_iterations))],
        ["Minimum Iterations per Path", int(np.min(all_iterations))],
    ]

    for idx, row_data in enumerate(overview_rows, 2):
        _write_row(ws4, idx, row_data, border, center_align, alternate_fill if idx % 2 == 0 else None)

    wb.save(filepath)
    print(f"\n Timing metrics exported to: {filepath}")
    return filepath


def run_single_experiment(run_num, max_iterations=3000):
    """Run one DQN + standard PSO ."""
    print(f"\n{'=' * 80}")
    print(f"Start run  {run_num}  run: DQN + standard PSO ")
    print(f"{'=' * 80}")

    run_start = time.perf_counter()
    path_results = []

    for path_idx, target_path in enumerate(target_paths):
        print(f"\nRun {run_num} run - Path {path_idx + 1}")
        path_total_start = time.perf_counter()

        # 1) DQN training
        dqn_train_start = time.perf_counter()
        agent = train_dqn_for_path(path_idx, target_path, num_samples=200)
        T_DQN_train = time.perf_counter() - dqn_train_start

        # 2) DQNSample Screening Time
        screening_start = time.perf_counter()
        dqn_samples = agent.get_best_samples(target_path, num_samples=20)
        T_sample_screening = time.perf_counter() - screening_start

        # 3) DQN
        direct_check_start = time.perf_counter()
        direct_solution_found = False
        for state_tuple, reward, sim, triggered in dqn_samples:
            if sim >= 1.0 or target_path.issubset(triggered):
                direct_solution_found = True
                break
        T_direct_check = time.perf_counter() - direct_check_start

        T_PSO_init = 0.0
        T_PSO_search = 0.0
        iterations_used = 0

        if direct_solution_found:
            print(
                f"  Path {path_idx + 1}: DQN | "
                f"DQN training {T_DQN_train:.6f}s | screening {T_sample_screening:.6f}s | iterations {iterations_used}"
            )
        else:
            # 4) standard PSOInitialization Time
            pso_init_start = time.perf_counter()
            pso = BasicPSO(target_path, swarm_size=20, dqn_samples=dqn_samples)
            T_PSO_init = time.perf_counter() - pso_init_start

            # 5) standard PSOSearch Time
            pso_search_start = time.perf_counter()
            iterations_used = max_iterations
            for iteration in range(max_iterations):
                pso.update()
                if pso.global_best_fitness >= 1.0:
                    iterations_used = iteration + 1
                    break
            T_PSO_search = time.perf_counter() - pso_search_start

            print(
                f"  Path {path_idx + 1}: "
                f"DQN training {T_DQN_train:.6f}s | screening {T_sample_screening:.6f}s | "
                f"PSO {T_PSO_init:.6f}s | PSO {T_PSO_search:.6f}s | "
                f"iterations {iterations_used}"
            )

        T_path_algorithm_total = time.perf_counter() - path_total_start

        path_results.append({
            "path_idx": path_idx,
            "T_DQN_train": T_DQN_train,
            "T_sample_screening": T_sample_screening,
            "T_direct_check": T_direct_check,
            "T_PSO_init": T_PSO_init,
            "T_PSO_search": T_PSO_search,
            "T_path_algorithm_total": T_path_algorithm_total,
            "iterations": iterations_used,
        })

    T_run_algorithm_total = time.perf_counter() - run_start
    print(f"\nRun {run_num} runcompleted | Total Algorithm Time {T_run_algorithm_total:.6f}s")
    return {
        "run_num": run_num,
        "paths": path_results,
        "T_run_algorithm_total": T_run_algorithm_total,
    }


def run_multiple_experiments(num_runs):
    print("\n" + "=" * 80)
    print(f"DQN + standard PSO - {num_runs}: Metric")
    print("=" * 80)
    print(f"Number of Target Paths: {len(target_paths)}")
    print("Statistics: DQN training, Sample Screening Time, PSOInitialization Time, PSOSearch Time, Path Iterations")
    print("=" * 80)

    all_run_results = []
    experiment_start = time.perf_counter()

    for run_num in range(1, num_runs + 1):
        run_result = run_single_experiment(run_num)
        all_run_results.append(run_result)

    experiment_total_time = time.perf_counter() - experiment_start
    print(f"\n{'=' * 80}")
    print(f"All {num_runs} runcompleted | Total Experiment Wall-clock Time {experiment_total_time:.6f}s")
    print(f"{'=' * 80}\n")
    return all_run_results, experiment_total_time


if __name__ == "__main__":
    print("=" * 80)
    print("DQN + standard PSO - MetricPath Iterations")
    print("=" * 80)
    print(f"Current configuration: Run = {NUM_RUNS}")
    print(f"Number of Paths: {len(target_paths)}")
    print(f"Device: {device}")
    print("=" * 80)

    if len(sys.argv) > 1:
        try:
            NUM_RUNS = int(sys.argv[1])
            print(f"Read from command line: Run = {NUM_RUNS}")
        except ValueError:
            print(f"Invalid command-line argument, using default number of runs {NUM_RUNS}")

    all_results, total_time = run_multiple_experiments(num_runs=NUM_RUNS)
    export_time_metrics_to_excel(all_results, total_time)
    print("\nProgram completed")
