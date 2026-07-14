import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

def execute_validation_rules(weather, time_period, z):
    """验证规则函数 - weather, time_period, z组合（删除覆盖率为0的分支后重新编号版本）"""

    triggered = set()

    # 将z映射到合适的范围以匹配条件逻辑
    x = z  # 直接使用z作为x
    y = (weather * time_period * 10 + z) % 100 + 1  # 基于输入参数计算y值

    # 1-7: 早高峰组合（time_period == 1）
    if time_period == 1:
        if (x < 60 and y > 75) != (x < 60 and y < 75):
            triggered.add(1)
        if (x > 60 and y > 70) != (x < 60 and y > 70):
            triggered.add(2)
        if (x < 50 and y < 40) != (x > 50 and y < 40):
            triggered.add(3)
        if (x > 78 and 45 < y < 70) != (x < 78 and 45 < y < 70):
            triggered.add(4)
        if (45 < x < 70 and y > 78) != (45 < x < 70 and y < 78):
            triggered.add(5)
        if (x < 55 and 50 < y < 75) != (x > 55 and 50 < y < 75):
            triggered.add(6)
        if (50 < x < 75 and y < 55) != (50 < x < 75 and y > 55):
            triggered.add(7)

    # 8-14: 晚高峰组合（time_period == 2）
    if time_period == 2:
        if (x < 60 and y > 75) != (x < 60 and y < 75):
            triggered.add(8)
        if (x > 60 and y > 70) != (x < 60 and y > 70):
            triggered.add(9)
        if (x < 55 and y < 45) != (x > 55 and y < 45):
            triggered.add(10)
        if (45 < x < 70 and y > 78) != (45 < x < 70 and y < 78):
            triggered.add(11)
        if (x > 78 and 45 < y < 70) != (x < 78 and 45 < y < 70):
            triggered.add(12)
        if (55 < x < 75 and y < 50) != (55 < x < 75 and y > 50):
            triggered.add(13)
        if (x < 50 and 55 < y < 75) != (x > 50 and 55 < y < 75):
            triggered.add(14)

    # 15-19: 午餐时间组合（time_period == 3）
    if time_period == 3:
        if (x > 60 and 40 < y < 65) != (x < 60 and 40 < y < 65):
            triggered.add(15)
        if (40 < x < 65 and y > 60) != (40 < x < 65 and y < 60):
            triggered.add(16)
        if (45 < x < 70 and 45 < y < 60) != (45 < x < 70 and y > 60):
            triggered.add(17)  # Fixed: 45 > y > 60 was invalid
        if (x < 50 and y < 40) != (x > 50 and y < 40):
            triggered.add(18)
        if (x > 65 and y < 45) != (x < 65 and y < 45):
            triggered.add(19)

    # 20-25: 夜间组合（time_period == 4）
    if time_period == 4:
        if (x < 45 and y < 35) != (x > 45 and y < 35):
            triggered.add(20)
        if (x > 60 and y < 40) != (x < 60 and y < 40):
            triggered.add(21)
        if (x < 50 and y > 70) != (x > 50 and y > 70):
            triggered.add(22)
        if (45 < x < 70 and 45 < y < 60) != (x > 70 and 45 < y < 60):
            triggered.add(23)  # Fixed: 45 > x > 70 was invalid
        if (x < 35 and y < 25) != (x > 35 and y < 25):
            triggered.add(24)
        if (40 < x < 65 and y < 45) != (40 < x < 65 and y > 45):
            triggered.add(25)

    # 26-28: 周末组合（time_period == 5）
    if time_period == 5:
        if (x < 60 and y < 50) != (x > 60 and y < 50):
            triggered.add(26)
        if (x > 65 and y > 75) != (x < 65 and y > 75):
            triggered.add(27)
        if (x > 60 and y < 45) != (x < 60 and y < 45):
            triggered.add(28)

    # 29-33: 假日组合（time_period == 6）
    if time_period == 6:
        if (40 < x < 70 and 40 < y < 60) != (x > 70 and 40 < y < 60):
            triggered.add(29)  # Fixed: 40 > x > 70 was invalid
        if (x < 55 and y < 45) != (x > 55 and y < 45):
            triggered.add(30)
        if (x > 60 and y < 50) != (x < 60 and y < 50):
            triggered.add(31)
        if (x < 60 and y > 70) != (x > 60 and y > 70):
            triggered.add(32)
        if (x > 65 and y > 75) != (x < 65 and y > 75):
            triggered.add(33)

    # 34-68: 天气相关扩展规则
    if weather == 1:  # 晴天
        if (time_period in [1, 2] and x > 70) != (time_period in [1, 2] and x < 70):
            triggered.add(34)
        if (time_period in [1, 2] and y > 70) != (time_period in [1, 2] and y < 70):
            triggered.add(35)
        if (time_period in [3, 4] and x < 50) != (time_period in [3, 4] and x > 50):
            triggered.add(36)
        if (time_period in [3, 4] and y < 50) != (time_period in [3, 4] and y > 50):
            triggered.add(37)
        if (time_period in [5, 6] and 40 < x < 80) != (time_period in [5, 6] and x > 80):
            triggered.add(38)  # Fixed: 40 > x > 80 was invalid
        if (time_period in [5, 6] and 40 < y < 80) != (time_period in [5, 6] and y > 80):
            triggered.add(39)  # Fixed: 40 > y > 80 was invalid

    if weather == 2:  # 雨天
        if (time_period in [1, 2] and x > 75) != (time_period in [1, 2] and x < 75):
            triggered.add(40)
        if (time_period in [1, 2] and y < 60) != (time_period in [1, 2] and y > 60):
            triggered.add(41)
        if (time_period in [3, 4] and x < 45) != (time_period in [3, 4] and x > 45):
            triggered.add(42)
        if (time_period in [3, 4] and y > 65) != (time_period in [3, 4] and y < 65):
            triggered.add(43)
        if (time_period in [5, 6] and 35 < x < 75) != (time_period in [5, 6] and x > 75):
            triggered.add(44)  # Fixed: 35 > x > 75 was invalid
        if (time_period in [5, 6] and 35 < y < 75) != (time_period in [5, 6] and y > 75):
            triggered.add(45)  # Fixed: 35 > y > 75 was invalid

    if weather == 3:  # 雾天
        if (time_period in [1, 2] and x > 60) != (time_period in [1, 2] and x < 60):
            triggered.add(46)
        if (time_period in [1, 2] and y > 65) != (time_period in [1, 2] and y < 65):
            triggered.add(47)
        if (time_period in [3, 4] and x < 55) != (time_period in [3, 4] and x > 55):
            triggered.add(48)
        if (time_period in [3, 4] and y < 55) != (time_period in [3, 4] and y > 55):
            triggered.add(49)
        if (time_period in [5, 6] and 30 < x < 70) != (time_period in [5, 6] and x > 70):
            triggered.add(50)  # Fixed: 30 > x > 70 was invalid
        if (time_period in [5, 6] and 30 < y < 70) != (time_period in [5, 6] and y > 70):
            triggered.add(51)  # Fixed: 30 > y > 70 was invalid

    if weather == 4:  # 雪天
        if (time_period in [1, 2] and x > 65) != (time_period in [1, 2] and x < 65):
            triggered.add(52)
        if (time_period in [1, 2] and y < 55) != (time_period in [1, 2] and y > 55):
            triggered.add(53)
        if (time_period in [3, 4] and x < 40) != (time_period in [3, 4] and x > 40):
            triggered.add(54)
        if (time_period in [3, 4] and y > 60) != (time_period in [3, 4] and y < 60):
            triggered.add(55)
        if (time_period in [5, 6] and 25 < x < 65) != (time_period in [5, 6] and x > 65):
            triggered.add(56)  # Fixed: 25 > x > 65 was invalid
        if (time_period in [5, 6] and 25 < y < 65) != (time_period in [5, 6] and y > 65):
            triggered.add(57)  # Fixed: 25 > y > 65 was invalid

    if weather == 5:  # 风天
        if (time_period in [1, 2] and x > 70) != (time_period in [1, 2] and x < 70):
            triggered.add(58)
        if (time_period in [1, 2] and y > 60) != (time_period in [1, 2] and y < 60):
            triggered.add(59)
        if (time_period in [3, 4] and x < 35) != (time_period in [3, 4] and x > 35):
            triggered.add(60)
        if (time_period in [3, 4] and y < 40) != (time_period in [3, 4] and y > 40):
            triggered.add(61)
        if (time_period in [5, 6] and 20 < x < 60) != (time_period in [5, 6] and x > 60):
            triggered.add(62)  # Fixed: 20 > x > 60 was invalid
        if (time_period in [5, 6] and 20 < y < 60) != (time_period in [5, 6] and y > 60):
            triggered.add(63)  # Fixed: 20 > y > 60 was invalid

    if weather == 6:  # 暴雨
        if (time_period in [1, 2] and x > 55) != (time_period in [1, 2] and x < 55):
            triggered.add(64)
        if (time_period in [1, 2] and y > 55) != (time_period in [1, 2] and y < 55):
            triggered.add(65)
        if (time_period in [3, 4] and x < 45) != (time_period in [3, 4] and x > 45):
            triggered.add(66)
        if (time_period in [3, 4] and y < 45) != (time_period in [3, 4] and y > 45):
            triggered.add(67)
        if (time_period in [5, 6] and 15 < x < 55) != (time_period in [5, 6] and x > 55):
            triggered.add(68)  # Fixed: 15 > x > 55 was invalid

    # 69-78: 复合条件（多参数组合）
    if weather + time_period > 6:
        if (x > 50 and y > 50) != (x < 50 and y > 50):
            triggered.add(69)
        if (x < 50 and y < 50) != (x > 50 and y < 50):
            triggered.add(70)
        if (x > y) != (x < y):
            triggered.add(71)
        if (x < y) != (x > y):
            triggered.add(72)
        if (abs(x - y) < 20) != (abs(x - y) > 20):
            triggered.add(73)

    if weather + time_period <= 6:
        if (x > 60 or y > 60) != (x < 60 or y > 60):
            triggered.add(74)
        if (x < 40 or y < 40) != (x > 40 or y < 40):
            triggered.add(75)
        if (x + y > 100) != (x + y < 100):
            triggered.add(76)
        if (x + y < 80) != (x + y > 80):
            triggered.add(77)
        if (abs(x - y) > 30) != (abs(x - y) < 30):
            triggered.add(78)

    # 79-88: 数值关系条件
    if weather % 2 == time_period % 2:  # 同奇偶性
        if (x % 10 < 5) != (x % 10 > 5):
            triggered.add(79)
        if (y % 10 >= 5) != (y % 10 < 5):
            triggered.add(80)
        if ((x + y) % 3 == 0) != ((x + y) % 3 == 1):
            triggered.add(81)
        if ((x * y) % 7 == 0) != ((x * y) % 7 == 1):
            triggered.add(82)
        if (x // 10 == y // 10) != (x // 10 != y // 10):
            triggered.add(83)

    if weather % 2 != time_period % 2:  # 不同奇偶性
        if (x > 75 or y > 75) != (x < 75 or y > 75):
            triggered.add(84)
        if (x < 25 or y < 25) != (x > 25 or y < 25):
            triggered.add(85)
        if (max(x, y) - min(x, y) > 40) != (max(x, y) - min(x, y) < 40):
            triggered.add(86)
        if ((x + y) // 2 > 50) != ((x + y) // 2 < 50):
            triggered.add(87)
        if (weather * time_period > 15) != (weather * time_period < 15):
            triggered.add(88)

    # 89-95: 高级组合条件（奇数天气）
    if weather in [1, 3, 5]:  # 奇数天气
        if (time_period in [1, 3, 5] and x > 40) != (time_period in [1, 3, 5] and x < 40):
            triggered.add(89)
        if (time_period in [2, 4, 6] and y > 40) != (time_period in [2, 4, 6] and y < 40):
            triggered.add(90)
        if (x % 20 < 10 and y % 20 < 10) != (x % 20 > 10 and y % 20 < 10):
            triggered.add(91)
        if (x + weather * 10 > 50) != (x + weather * 10 < 50):
            triggered.add(92)
        if (y + time_period * 10 > 50) != (y + time_period * 10 < 50):
            triggered.add(93)
        if (time_period in [1, 3, 5] and x < 60) != (time_period in [1, 3, 5] and x > 60):
            triggered.add(94)
        if (time_period in [2, 4, 6] and y < 60) != (time_period in [2, 4, 6] and y > 60):
            triggered.add(95)

    # 96-98: 偶数天气条件
    if weather in [2, 4, 6]:  # 偶数天气
        if ((x + y) % weather == 0) != ((x + y) % weather == 1):
            triggered.add(96)
        if (x * weather > 100) != (x * weather < 100):
            triggered.add(97)
        if (y * time_period > 100) != (y * time_period < 100):
            triggered.add(98)

    # 99-100: 最后的复杂条件
    if ((weather * time_period + z) % 7 == 0) != ((weather * time_period + z) % 7 == 1):
        triggered.add(99)
    if (max(weather, time_period) * min(x, y) > 150) != (max(weather, time_period) * min(x, y) < 150):
        triggered.add(100)

    return triggered

def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """"""
    generated_path = execute_validation_rules(particle[0], particle[1], particle[2])

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # : x:1-50, y:1-50, z:1-50
        self.bounds = bounds if bounds else [(1, 100), (1, 100), (1, 100)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """target pathsPSO"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(particles[i][0], particles[i][1], particles[i][2]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(gbest_particle[0], gbest_particle[1], gbest_particle[2]),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """Path PSO"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - Path ")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations")
    print(f"Path : {len(target_paths)}")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"Path {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "" if result['success'] else f"({result['best_fitness']:.3f})"
        print(f"{status} | {result['time']:.2f}s | iterations{result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f": {success_count}/{len(target_paths)} ({success_rate:.1f}%) | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs}")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations, {len(target_paths)}Path ")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"---  {run_idx}/{num_runs} ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f": {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs} runcompleted | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 1: 
    ws1 = wb.active
    ws1.title = ""
    ws1.sheet_view.showGridLines = False

    headers = ["Run", "", "", "", "Average Iterations", "(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f" {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # 2: Path 
    ws2 = wb.create_sheet(title="Path ")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Path ID", "", "", "", "Average Iterations", "Minimum Iterations", "Maximum Iterations"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"Path  {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # 3: 
    ws3 = wb.create_sheet(title="")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Path ", "", "(x,y,z)", "", "Iterations", "Path "]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"{run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # 4: target paths
    ws4 = wb.create_sheet(title="target paths")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Path ID", "target paths", ""]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"Path  {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f" : {filename}")
    print(f"{'=' * 70}")
    print(f":")
    print(f"  1.        - {len(all_results)} run")
    print(f"  2. Path        - Path ")
    print(f"  3.    -  runPath ")
    print(f"  4. target paths       - target paths")
    print(f"{'=' * 70}\n")

    return filename


def main():


    target_paths = [
        {15, 16, 48, 49, 74, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 91, 92, 93, 94, 99, 100},
        {16, 18, 19, 60, 61, 70, 71, 72, 73, 80, 81, 82, 83, 89, 91, 92, 93, 94, 99, 100},
        {1, 4, 6, 46, 47, 74, 75, 76, 77, 78, 80, 81, 82, 83, 89, 92, 93, 94, 99, 100},
        {30, 31, 50, 51, 70, 71, 72, 73, 84, 85, 86, 87, 88, 91, 92, 93, 95, 99, 100},
        {18, 19, 36, 37, 74, 76, 77, 78, 79, 80, 81, 82, 83, 89, 92, 93, 94, 99, 100},
        {20, 24, 25, 36, 37, 76, 77, 78, 84, 86, 87, 88, 90, 91, 92, 93, 95, 99, 100},
        {8, 12, 34, 35, 74, 75, 76, 77, 78, 84, 86, 87, 88, 90, 91, 92, 93, 95, 100},
        {8, 10, 58, 59, 70, 71, 72, 73, 84, 85, 86, 87, 88, 91, 92, 93, 95, 99, 100},
        {8, 14, 46, 47, 75, 76, 77, 78, 84, 85, 86, 87, 88, 90, 92, 93, 95, 99, 100},
        {1, 2, 6, 46, 47, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 92, 93, 94, 100},
        {39, 74, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 91, 92, 93, 94, 99, 100},
        {20, 21, 60, 61, 70, 71, 72, 73, 84, 85, 86, 87, 88, 90, 92, 93, 95, 99},
        {8, 9, 11, 13, 40, 41, 75, 76, 77, 78, 79, 80, 81, 83, 96, 97, 98, 100},
        {18, 19, 54, 55, 70, 71, 72, 73, 84, 86, 87, 88, 96, 97, 98, 99, 100},
        {27, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 91, 92, 93, 94, 99, 100},
        {25, 48, 49, 69, 71, 72, 73, 84, 85, 86, 87, 88, 90, 92, 93, 95, 100},
        {26, 28, 62, 70, 71, 72, 73, 80, 81, 82, 83, 89, 91, 92, 93, 94, 100},
        {32, 33, 68, 69, 71, 72, 73, 79, 80, 81, 82, 83, 96, 97, 98, 99, 100},
        {1, 52, 53, 74, 75, 76, 77, 78, 84, 85, 86, 87, 88, 97, 98, 99, 100},
        {8, 12, 14, 64, 65, 69, 71, 72, 73, 80, 81, 82, 83, 96, 97, 98, 100},
        {1, 3, 64, 65, 70, 71, 72, 73, 84, 86, 87, 88, 96, 97, 98, 99, 100},
        {22, 36, 37, 76, 77, 78, 85, 86, 87, 88, 90, 91, 93, 95, 100},
        {31, 45, 70, 71, 72, 73, 79, 80, 81, 83, 96, 97, 98, 99, 100},
        {22, 66, 67, 69, 71, 72, 73, 79, 80, 82, 83, 97, 98, 100},
        {44, 45, 69, 71, 72, 73, 79, 80, 83, 96, 97, 98, 99, 100},
        {57, 71, 72, 73, 79, 80, 83, 97, 98, 100},
        {15, 16, 17, 48, 49, 74, 75, 76, 77, 78, 79, 80, 82, 83, 89, 91, 92, 93, 94, 100},
        {1, 2, 5, 46, 47, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 91, 92, 93, 94, 100},
        {20, 21, 25, 42, 43, 74, 76, 77, 78, 79, 80, 81, 82, 83, 96, 97, 98, 99, 100},
        {2, 5, 7, 40, 41, 75, 76, 77, 78, 84, 85, 86, 87, 88, 96, 97, 98, 99, 100},
        {26, 28, 56, 57, 70, 71, 72, 73, 84, 85, 86, 87, 88, 96, 97, 98, 99, 100},
        {26, 28, 38, 74, 76, 77, 78, 80, 81, 82, 83, 89, 91, 92, 93, 94, 100},
        {30, 31, 62, 63, 70, 71, 72, 73, 84, 86, 87, 88, 90, 91, 92, 93, 95},
        {29, 62, 63, 71, 72, 73, 84, 85, 86, 87, 88, 90, 92, 93, 95, 100},
        {23, 25, 60, 61, 71, 72, 73, 84, 85, 86, 87, 88, 90, 92, 93, 95, 100}
    ]

    print("=" * 70)
    print("baseline PSO")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("Program completed")


if __name__ == "__main__":
    main()