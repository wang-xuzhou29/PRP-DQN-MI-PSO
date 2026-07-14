import os
import random
from collections import deque
import numpy as np
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# --- 全局状态范围配置 ---
STATE_MIN_X, STATE_MAX_X = 1, 6
STATE_MIN_Y, STATE_MAX_Y = 1, 6
STATE_MIN_Z, STATE_MAX_Z = 1, 6

def generate_input():
    return [
        random.randint(STATE_MIN_X, STATE_MAX_X),
        random.randint(STATE_MIN_Y, STATE_MAX_Y),
        random.randint(STATE_MIN_Z, STATE_MAX_Z)
    ]


def execute_Tr(weather, time_period, z):
    """验证规则函数 - weather, time_period, z组合（删除覆盖率为0的分支后重新编号版本）"""

    triggered = set()

    # 将z映射到合适的范围以匹配条件逻辑
    x = z  # 直接使用z作为x
    y = (weather * time_period * 10 + z) % 100 + 1  # 基于输入参数计算y值

    # 1-7: 早高峰组合（time_period == 1）
    if time_period == 1:
        if (x < 60 and y > 75) != (x < 60 and y < 75):
            triggered.add(1)
        if (x > 60 and y > 70) != (x < 60 and y > 70):
            triggered.add(2)
        if (x < 50 and y < 40) != (x > 50 and y < 40):
            triggered.add(3)
        if (x > 78 and 45 < y < 70) != (x < 78 and 45 < y < 70):
            triggered.add(4)
        if (45 < x < 70 and y > 78) != (45 < x < 70 and y < 78):
            triggered.add(5)
        if (x < 55 and 50 < y < 75) != (x > 55 and 50 < y < 75):
            triggered.add(6)
        if (50 < x < 75 and y < 55) != (50 < x < 75 and y > 55):
            triggered.add(7)

    # 8-14: 晚高峰组合（time_period == 2）
    if time_period == 2:
        if (x < 60 and y > 75) != (x < 60 and y < 75):
            triggered.add(8)
        if (x > 60 and y > 70) != (x < 60 and y > 70):
            triggered.add(9)
        if (x < 55 and y < 45) != (x > 55 and y < 45):
            triggered.add(10)
        if (45 < x < 70 and y > 78) != (45 < x < 70 and y < 78):
            triggered.add(11)
        if (x > 78 and 45 < y < 70) != (x < 78 and 45 < y < 70):
            triggered.add(12)
        if (55 < x < 75 and y < 50) != (55 < x < 75 and y > 50):
            triggered.add(13)
        if (x < 50 and 55 < y < 75) != (x > 50 and 55 < y < 75):
            triggered.add(14)

    # 15-19: 午餐时间组合（time_period == 3）
    if time_period == 3:
        if (x > 60 and 40 < y < 65) != (x < 60 and 40 < y < 65):
            triggered.add(15)
        if (40 < x < 65 and y > 60) != (40 < x < 65 and y < 60):
            triggered.add(16)
        if (45 < x < 70 and 45 < y < 60) != (45 < x < 70 and y > 60):
            triggered.add(17)  # Fixed: 45 > y > 60 was invalid
        if (x < 50 and y < 40) != (x > 50 and y < 40):
            triggered.add(18)
        if (x > 65 and y < 45) != (x < 65 and y < 45):
            triggered.add(19)

    # 20-25: 夜间组合（time_period == 4）
    if time_period == 4:
        if (x < 45 and y < 35) != (x > 45 and y < 35):
            triggered.add(20)
        if (x > 60 and y < 40) != (x < 60 and y < 40):
            triggered.add(21)
        if (x < 50 and y > 70) != (x > 50 and y > 70):
            triggered.add(22)
        if (45 < x < 70 and 45 < y < 60) != (x > 70 and 45 < y < 60):
            triggered.add(23)  # Fixed: 45 > x > 70 was invalid
        if (x < 35 and y < 25) != (x > 35 and y < 25):
            triggered.add(24)
        if (40 < x < 65 and y < 45) != (40 < x < 65 and y > 45):
            triggered.add(25)

    # 26-28: 周末组合（time_period == 5）
    if time_period == 5:
        if (x < 60 and y < 50) != (x > 60 and y < 50):
            triggered.add(26)
        if (x > 65 and y > 75) != (x < 65 and y > 75):
            triggered.add(27)
        if (x > 60 and y < 45) != (x < 60 and y < 45):
            triggered.add(28)

    # 29-33: 假日组合（time_period == 6）
    if time_period == 6:
        if (40 < x < 70 and 40 < y < 60) != (x > 70 and 40 < y < 60):
            triggered.add(29)  # Fixed: 40 > x > 70 was invalid
        if (x < 55 and y < 45) != (x > 55 and y < 45):
            triggered.add(30)
        if (x > 60 and y < 50) != (x < 60 and y < 50):
            triggered.add(31)
        if (x < 60 and y > 70) != (x > 60 and y > 70):
            triggered.add(32)
        if (x > 65 and y > 75) != (x < 65 and y > 75):
            triggered.add(33)

    # 34-68: 天气相关扩展规则
    if weather == 1:  # 晴天
        if (time_period in [1, 2] and x > 70) != (time_period in [1, 2] and x < 70):
            triggered.add(34)
        if (time_period in [1, 2] and y > 70) != (time_period in [1, 2] and y < 70):
            triggered.add(35)
        if (time_period in [3, 4] and x < 50) != (time_period in [3, 4] and x > 50):
            triggered.add(36)
        if (time_period in [3, 4] and y < 50) != (time_period in [3, 4] and y > 50):
            triggered.add(37)
        if (time_period in [5, 6] and 40 < x < 80) != (time_period in [5, 6] and x > 80):
            triggered.add(38)  # Fixed: 40 > x > 80 was invalid
        if (time_period in [5, 6] and 40 < y < 80) != (time_period in [5, 6] and y > 80):
            triggered.add(39)  # Fixed: 40 > y > 80 was invalid

    if weather == 2:  # 雨天
        if (time_period in [1, 2] and x > 75) != (time_period in [1, 2] and x < 75):
            triggered.add(40)
        if (time_period in [1, 2] and y < 60) != (time_period in [1, 2] and y > 60):
            triggered.add(41)
        if (time_period in [3, 4] and x < 45) != (time_period in [3, 4] and x > 45):
            triggered.add(42)
        if (time_period in [3, 4] and y > 65) != (time_period in [3, 4] and y < 65):
            triggered.add(43)
        if (time_period in [5, 6] and 35 < x < 75) != (time_period in [5, 6] and x > 75):
            triggered.add(44)  # Fixed: 35 > x > 75 was invalid
        if (time_period in [5, 6] and 35 < y < 75) != (time_period in [5, 6] and y > 75):
            triggered.add(45)  # Fixed: 35 > y > 75 was invalid

    if weather == 3:  # 雾天
        if (time_period in [1, 2] and x > 60) != (time_period in [1, 2] and x < 60):
            triggered.add(46)
        if (time_period in [1, 2] and y > 65) != (time_period in [1, 2] and y < 65):
            triggered.add(47)
        if (time_period in [3, 4] and x < 55) != (time_period in [3, 4] and x > 55):
            triggered.add(48)
        if (time_period in [3, 4] and y < 55) != (time_period in [3, 4] and y > 55):
            triggered.add(49)
        if (time_period in [5, 6] and 30 < x < 70) != (time_period in [5, 6] and x > 70):
            triggered.add(50)  # Fixed: 30 > x > 70 was invalid
        if (time_period in [5, 6] and 30 < y < 70) != (time_period in [5, 6] and y > 70):
            triggered.add(51)  # Fixed: 30 > y > 70 was invalid

    if weather == 4:  # 雪天
        if (time_period in [1, 2] and x > 65) != (time_period in [1, 2] and x < 65):
            triggered.add(52)
        if (time_period in [1, 2] and y < 55) != (time_period in [1, 2] and y > 55):
            triggered.add(53)
        if (time_period in [3, 4] and x < 40) != (time_period in [3, 4] and x > 40):
            triggered.add(54)
        if (time_period in [3, 4] and y > 60) != (time_period in [3, 4] and y < 60):
            triggered.add(55)
        if (time_period in [5, 6] and 25 < x < 65) != (time_period in [5, 6] and x > 65):
            triggered.add(56)  # Fixed: 25 > x > 65 was invalid
        if (time_period in [5, 6] and 25 < y < 65) != (time_period in [5, 6] and y > 65):
            triggered.add(57)  # Fixed: 25 > y > 65 was invalid

    if weather == 5:  # 风天
        if (time_period in [1, 2] and x > 70) != (time_period in [1, 2] and x < 70):
            triggered.add(58)
        if (time_period in [1, 2] and y > 60) != (time_period in [1, 2] and y < 60):
            triggered.add(59)
        if (time_period in [3, 4] and x < 35) != (time_period in [3, 4] and x > 35):
            triggered.add(60)
        if (time_period in [3, 4] and y < 40) != (time_period in [3, 4] and y > 40):
            triggered.add(61)
        if (time_period in [5, 6] and 20 < x < 60) != (time_period in [5, 6] and x > 60):
            triggered.add(62)  # Fixed: 20 > x > 60 was invalid
        if (time_period in [5, 6] and 20 < y < 60) != (time_period in [5, 6] and y > 60):
            triggered.add(63)  # Fixed: 20 > y > 60 was invalid

    if weather == 6:  # 暴雨
        if (time_period in [1, 2] and x > 55) != (time_period in [1, 2] and x < 55):
            triggered.add(64)
        if (time_period in [1, 2] and y > 55) != (time_period in [1, 2] and y < 55):
            triggered.add(65)
        if (time_period in [3, 4] and x < 45) != (time_period in [3, 4] and x > 45):
            triggered.add(66)
        if (time_period in [3, 4] and y < 45) != (time_period in [3, 4] and y > 45):
            triggered.add(67)
        if (time_period in [5, 6] and 15 < x < 55) != (time_period in [5, 6] and x > 55):
            triggered.add(68)  # Fixed: 15 > x > 55 was invalid

    # 69-78: 复合条件（多参数组合）
    if weather + time_period > 6:
        if (x > 50 and y > 50) != (x < 50 and y > 50):
            triggered.add(69)
        if (x < 50 and y < 50) != (x > 50 and y < 50):
            triggered.add(70)
        if (x > y) != (x < y):
            triggered.add(71)
        if (x < y) != (x > y):
            triggered.add(72)
        if (abs(x - y) < 20) != (abs(x - y) > 20):
            triggered.add(73)

    if weather + time_period <= 6:
        if (x > 60 or y > 60) != (x < 60 or y > 60):
            triggered.add(74)
        if (x < 40 or y < 40) != (x > 40 or y < 40):
            triggered.add(75)
        if (x + y > 100) != (x + y < 100):
            triggered.add(76)
        if (x + y < 80) != (x + y > 80):
            triggered.add(77)
        if (abs(x - y) > 30) != (abs(x - y) < 30):
            triggered.add(78)

    # 79-88: 数值关系条件
    if weather % 2 == time_period % 2:  # 同奇偶性
        if (x % 10 < 5) != (x % 10 > 5):
            triggered.add(79)
        if (y % 10 >= 5) != (y % 10 < 5):
            triggered.add(80)
        if ((x + y) % 3 == 0) != ((x + y) % 3 == 1):
            triggered.add(81)
        if ((x * y) % 7 == 0) != ((x * y) % 7 == 1):
            triggered.add(82)
        if (x // 10 == y // 10) != (x // 10 != y // 10):
            triggered.add(83)

    if weather % 2 != time_period % 2:  # 不同奇偶性
        if (x > 75 or y > 75) != (x < 75 or y > 75):
            triggered.add(84)
        if (x < 25 or y < 25) != (x > 25 or y < 25):
            triggered.add(85)
        if (max(x, y) - min(x, y) > 40) != (max(x, y) - min(x, y) < 40):
            triggered.add(86)
        if ((x + y) // 2 > 50) != ((x + y) // 2 < 50):
            triggered.add(87)
        if (weather * time_period > 15) != (weather * time_period < 15):
            triggered.add(88)

    # 89-95: 高级组合条件（奇数天气）
    if weather in [1, 3, 5]:  # 奇数天气
        if (time_period in [1, 3, 5] and x > 40) != (time_period in [1, 3, 5] and x < 40):
            triggered.add(89)
        if (time_period in [2, 4, 6] and y > 40) != (time_period in [2, 4, 6] and y < 40):
            triggered.add(90)
        if (x % 20 < 10 and y % 20 < 10) != (x % 20 > 10 and y % 20 < 10):
            triggered.add(91)
        if (x + weather * 10 > 50) != (x + weather * 10 < 50):
            triggered.add(92)
        if (y + time_period * 10 > 50) != (y + time_period * 10 < 50):
            triggered.add(93)
        if (time_period in [1, 3, 5] and x < 60) != (time_period in [1, 3, 5] and x > 60):
            triggered.add(94)
        if (time_period in [2, 4, 6] and y < 60) != (time_period in [2, 4, 6] and y > 60):
            triggered.add(95)

    # 96-98: 偶数天气条件
    if weather in [2, 4, 6]:  # 偶数天气
        if ((x + y) % weather == 0) != ((x + y) % weather == 1):
            triggered.add(96)
        if (x * weather > 100) != (x * weather < 100):
            triggered.add(97)
        if (y * time_period > 100) != (y * time_period < 100):
            triggered.add(98)

    # 99-100: 最后的复杂条件
    if ((weather * time_period + z) % 7 == 0) != ((weather * time_period + z) % 7 == 1):
        triggered.add(99)
    if (max(weather, time_period) * min(x, y) > 150) != (max(weather, time_period) * min(x, y) < 150):
        triggered.add(100)

    return triggered


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0


# targetPaths - now as sets (already correct)
TARGET_PATHS = [
    {15, 16, 48, 49, 74, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 91, 92, 93, 94, 99, 100},
    {16, 18, 19, 60, 61, 70, 71, 72, 73, 80, 81, 82, 83, 89, 91, 92, 93, 94, 99, 100},
    {1, 4, 6, 46, 47, 74, 75, 76, 77, 78, 80, 81, 82, 83, 89, 92, 93, 94, 99, 100},
    {30, 31, 50, 51, 70, 71, 72, 73, 84, 85, 86, 87, 88, 91, 92, 93, 95, 99, 100},
    {18, 19, 36, 37, 74, 76, 77, 78, 79, 80, 81, 82, 83, 89, 92, 93, 94, 99, 100},
    {20, 24, 25, 36, 37, 76, 77, 78, 84, 86, 87, 88, 90, 91, 92, 93, 95, 99, 100},
    {8, 12, 34, 35, 74, 75, 76, 77, 78, 84, 86, 87, 88, 90, 91, 92, 93, 95, 100},
    {8, 10, 58, 59, 70, 71, 72, 73, 84, 85, 86, 87, 88, 91, 92, 93, 95, 99, 100},
    {8, 14, 46, 47, 75, 76, 77, 78, 84, 85, 86, 87, 88, 90, 92, 93, 95, 99, 100},
    {1, 2, 6, 46, 47, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 92, 93, 94, 100},
    {39, 74, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 91, 92, 93, 94, 99, 100},
    {20, 21, 60, 61, 70, 71, 72, 73, 84, 85, 86, 87, 88, 90, 92, 93, 95, 99},
    {8, 9, 11, 13, 40, 41, 75, 76, 77, 78, 79, 80, 81, 83, 96, 97, 98, 100},
    {18, 19, 54, 55, 70, 71, 72, 73, 84, 86, 87, 88, 96, 97, 98, 99, 100},
    {27, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 91, 92, 93, 94, 99, 100},
    {25, 48, 49, 69, 71, 72, 73, 84, 85, 86, 87, 88, 90, 92, 93, 95, 100},
    {26, 28, 62, 70, 71, 72, 73, 80, 81, 82, 83, 89, 91, 92, 93, 94, 100},
    {32, 33, 68, 69, 71, 72, 73, 79, 80, 81, 82, 83, 96, 97, 98, 99, 100},
    {1, 52, 53, 74, 75, 76, 77, 78, 84, 85, 86, 87, 88, 97, 98, 99, 100},
    {8, 12, 14, 64, 65, 69, 71, 72, 73, 80, 81, 82, 83, 96, 97, 98, 100},
    {1, 3, 64, 65, 70, 71, 72, 73, 84, 86, 87, 88, 96, 97, 98, 99, 100},
    {22, 36, 37, 76, 77, 78, 85, 86, 87, 88, 90, 91, 93, 95, 100},
    {31, 45, 70, 71, 72, 73, 79, 80, 81, 83, 96, 97, 98, 99, 100},
    {22, 66, 67, 69, 71, 72, 73, 79, 80, 82, 83, 97, 98, 100},
    {44, 45, 69, 71, 72, 73, 79, 80, 83, 96, 97, 98, 99, 100},
    {57, 71, 72, 73, 79, 80, 83, 97, 98, 100},
    {15, 16, 17, 48, 49, 74, 75, 76, 77, 78, 79, 80, 82, 83, 89, 91, 92, 93, 94, 100},
    {1, 2, 5, 46, 47, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 91, 92, 93, 94, 100},
    {20, 21, 25, 42, 43, 74, 76, 77, 78, 79, 80, 81, 82, 83, 96, 97, 98, 99, 100},
    {2, 5, 7, 40, 41, 75, 76, 77, 78, 84, 85, 86, 87, 88, 96, 97, 98, 99, 100},
    {26, 28, 56, 57, 70, 71, 72, 73, 84, 85, 86, 87, 88, 96, 97, 98, 99, 100},
    {26, 28, 38, 74, 76, 77, 78, 80, 81, 82, 83, 89, 91, 92, 93, 94, 100},
    {30, 31, 62, 63, 70, 71, 72, 73, 84, 86, 87, 88, 90, 91, 92, 93, 95},
    {29, 62, 63, 71, 72, 73, 84, 85, 86, 87, 88, 90, 92, 93, 95, 100},
    {23, 25, 60, 61, 71, 72, 73, 84, 85, 86, 87, 88, 90, 92, 93, 95, 100}
]


class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 2000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(TARGET_PATHS)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }


def compute_robustness(state, path):
    base = execute_Tr(*state)
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                    np.clip(state[0] + dx, STATE_MIN_X, STATE_MAX_X),
                    np.clip(state[1] + dy, STATE_MIN_Y, STATE_MAX_Y),
                    np.clip(state[2] + dz, STATE_MIN_Z, STATE_MAX_Z)
                ])
                n_trig = execute_Tr(*neighbor)
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = TARGET_PATHS[target_path_idx]
    samples = []
    attempts = 0

    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(STATE_MIN_X, STATE_MAX_X),
            random.randint(STATE_MIN_Y, STATE_MAX_Y),
            random.randint(STATE_MIN_Z, STATE_MAX_Z)
        ])
        triggered = execute_Tr(*state)

        if not triggered:
            continue

        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)

        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }

        samples.append(sample_data)

    return samples


def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample[
                'robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)

        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]

    return selected_samples


def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = TARGET_PATHS[target_path_idx]
        samples = []
        attempts = 0

        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(STATE_MIN_X, STATE_MAX_X),
                random.randint(STATE_MIN_Y, STATE_MAX_Y),
                random.randint(STATE_MIN_Z, STATE_MAX_Z)
            ])
            triggered = execute_Tr(*state)

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)

            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)

        random.shuffle(samples)
        return samples[:config.top_k_samples]

    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")

        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)


def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}

    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates

    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}

        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config,
                                                         shared_candidates[path_idx])

            strategy_results[path_idx] = samples

        results[strategy_name] = strategy_results

    return results


def analyze_fitness_values(results, config):
    analysis_results = {}

    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []

        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])

        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }

        all_scores = []

        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])

        scores_array = np.array(all_scores)

        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)

        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })

        analysis_results[strategy_name] = analysis

    return analysis_results


def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []

    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)

    df = pd.DataFrame(df_data)
    return df, analysis_results


def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []

    print(f"Starting {num_runs} experiments...")

    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")

        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)

        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)

        print(f"Experiment {run_idx + 1} completed.")

    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df


def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')

    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width

    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []

    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]

        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)

    stats_df = pd.DataFrame(stats_data)

    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Results saved to: {output_path}")


def main():
    results_df = run_multiple_experiments(num_runs=20)

    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")

    save_results_to_excel(results_df, output_path)

    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)

    return results_df, output_path


if __name__ == "__main__":
    results, output_file = main()