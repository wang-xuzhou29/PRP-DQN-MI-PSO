import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

def execute_validation_rules(x, y, z):
    triggered = set()
    actions = []

    devices = {
        'main_light': 'green',
        'side_light': 'red',
        'pedestrian_light': 'red',
        'warning_system': 'off'
    }

    # Fixed syntax: properly formatted if statements
    if (x > 85 and y < 40 and z < 25) != (x > 90 and y < 40 and z < 25):
        triggered.add(1)
    if (x > 85 and y < 40 and z < 25) != (x > 85 and y < 35 and z < 25):
        triggered.add(2)
    if (x > 85 and y < 40 and z < 25) != (x > 85 and y < 40 and z < 20):
        triggered.add(3)
    if (x > 85 and y < 40 and z < 25) != (x > 80 and y < 40 and z < 25):
        triggered.add(4)
    if (x > 80 and y < 45 and z > 40) != (x > 80 or y < 45 and z > 40):
        triggered.add(5)
    if (x > 80 and y < 45 and z > 40) != (x > 80 and y < 45 or z > 40):
        triggered.add(6)
    if (x > 80 and y < 45 and z > 40) != (x > 60 and y < 45 and z > 40):
        triggered.add(7)
    if (x > 80 and y < 45 and z > 40) != (x > 80 and y < 25 and z > 40):
        triggered.add(8)
    if (x > 80 and y < 45 and z > 40) != (x > 80 and y < 45 and z > 20):
        triggered.add(9)
    if (x > 92 and y < 30 and z < 15) != (x > 92 or y < 30 and z < 15):
        triggered.add(10)
    if (x > 92 and y < 30 and z < 15) != (x > 92 and y < 30 or z < 15):
        triggered.add(11)
    if (x > 92 and y < 30 and z < 15) != (x > 72 and y < 30 and z < 15):
        triggered.add(12)
    if (x > 92 and y < 30 and z < 15) != (x > 92 and y < 10 and z < 15):
        triggered.add(13)
    if (x > 75 and 45 < y < 65 and z > 50) != (x > 55 and 45 < y < 65 and z > 50):
        triggered.add(14)
    if (x > 75 and 45 < y < 65 and z > 50) != (x > 75 and 35 < y < 65 and z > 50):
        triggered.add(15)
    if (x > 75 and 45 < y < 65 and z > 50) != (x > 75 and 45 < y < 55 and z > 50):
        triggered.add(16)
    if (x > 75 and 45 < y < 65 and z > 50) != (x > 75 and 45 < y < 65 and z > 40):
        triggered.add(17)
    if (x > 75 and 45 < y < 65 and z > 50) != (x > 75 and 45 < y < 65 or z > 50):
        triggered.add(18)
    if (x < 50 and y > 80 and z < 25) != (x < 50 or y > 80 and z < 25):
        triggered.add(19)
    if (x < 50 and y > 80 and z < 25) != (x < 50 and y > 80 or z < 25):
        triggered.add(20)
    if (x < 50 and y > 80 and z < 25) != (x < 25 and y > 80 and z < 25):
        triggered.add(21)
    if (x < 50 and y > 80 and z < 25) != (x < 50 and y > 40 and z < 25):
        triggered.add(22)
    if (x < 50 and y > 80 and z < 25) != (x < 50 and y > 80 and z < 15):
        triggered.add(23)
    if (x < 30 and y > 92 and z < 15) != (x < 30 or y > 92 and z < 15):
        triggered.add(24)
    if (x < 30 and y > 92 and z < 15) != (x < 30 and y > 92 or z < 15):
        triggered.add(25)
    if (x < 30 and y > 92 and z < 15) != (x < 70 and y > 92 and z < 15):
        triggered.add(26)
    if (x < 30 and y > 92 and z < 15) != (x < 30 and y > 42 and z < 15):
        triggered.add(27)
    if (x < 30 and y > 92 and z < 15) != (x < 30 and y > 92 and z < 5):
        triggered.add(28)
    if (x < 30 and y > 92 and z < 15) != (x < 30 and y > 62 and z < 15):
        triggered.add(29)
    if (x > 70 and y > 70 and z > 45) != (x > 70 or y > 70 and z > 45):
        triggered.add(30)
    if (x > 70 and y > 70 and z > 45) != (x > 70 and y > 70 or z > 45):
        triggered.add(31)
    if (x > 70 and y > 70 and z > 45) != (x > 50 and y > 70 and z > 45):
        triggered.add(32)
    if (x > 70 and y > 70 and z > 45) != (x > 70 and y > 50 and z > 45):
        triggered.add(33)
    if (x > 70 and y > 70 and z > 45) != (x > 70 and y > 70 and z > 25):
        triggered.add(34)
    if (x > 70 and y > 70 and z > 45) != (x > 35 and y > 70 and z > 45):
        triggered.add(35)
    if (x > 88 and y > 88 and 25 < z < 45) != (x > 88 or y > 88 and 25 < z < 45):
        triggered.add(36)
    if (x > 88 and y > 88 and 25 < z < 45) != (x > 88 and y > 88 or 25 < z < 45):
        triggered.add(37)
    if (x > 88 and y > 88 and 25 < z < 45) != (x > 44 and y > 88 and 25 < z < 45):
        triggered.add(38)
    if (x > 88 and y > 88 and 25 < z < 45) != (x > 88 and y > 44 and 25 < z < 45):
        triggered.add(39)
    if (x > 88 and y > 88 and 25 < z < 45) != (x > 88 and y > 88 and 15 < z < 45):
        triggered.add(40)
    if (x > 88 and y > 88 and 25 < z < 45) != (x > 88 and y > 88 and 25 < z < 25):
        triggered.add(41)
    if (x > 75 and y > 75 and z > 55) != (x > 35 and y > 75 and z > 55):
        triggered.add(42)
    if (x > 75 and y > 75 and z > 55) != (x > 75 or y > 75 and z > 55):
        triggered.add(43)
    if (x > 75 and y > 75 and z > 55) != (x > 75 and y > 75 or z > 55):
        triggered.add(44)
    if (x > 75 and y > 75 and z > 55) != (x > 75 and y > 35 and z > 55):
        triggered.add(45)
    if (x > 75 and y > 75 and z > 55) != (x > 75 and y > 75 and z > 25):
        triggered.add(46)
    if (x < 40 and y < 40 and z > 40) != (x < 40 or y < 40 and z > 40):
        triggered.add(47)
    if (x < 40 and y < 40 and z > 40) != (x < 40 and y < 40 or z > 40):
        triggered.add(48)
    if (x < 40 and y < 40 and z > 40) != (x < 20 and y < 40 and z > 40):
        triggered.add(49)
    if (x < 40 and y < 40 and z > 40) != (x < 40 and y < 20 and z > 40):
        triggered.add(50)
    if (x < 40 and y < 40 and z > 40) != (x < 40 and y < 40 and z > 20):
        triggered.add(51)
    if (x < 25 and y < 25 and 20 < z < 40) != (x < 25 or y < 25 and 20 < z < 40):
        triggered.add(52)
    if (x < 25 and y < 25 and 20 < z < 40) != (x < 25 and y < 25 or 20 < z < 40):
        triggered.add(53)
    if (x < 25 and y < 25 and 20 < z < 40) != (x < 15 and y < 25 and 20 < z < 40):
        triggered.add(54)
    if (x < 25 and y < 25 and 20 < z < 40) != (x < 5 and y < 25 and 20 < z < 40):
        triggered.add(55)
    if (x < 25 and y < 25 and 20 < z < 40) != (x < 25 and y < 15 and 20 < z < 40):
        triggered.add(56)
    if (x < 25 and y < 25 and 20 < z < 40) != (x < 25 and y < 5 and 20 < z < 40):
        triggered.add(57)
    if (x < 25 and y < 25 and 20 < z < 40) != (x < 25 and y < 25 and 20 < z < 20):
        triggered.add(58)
    if (x < 25 and y < 25 and 20 < z < 40) != (x < 25 and y < 25 and 20 < z < 10):
        triggered.add(59)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (45 < x < 70 or 45 < y < 70 and 25 < z < 45):
        triggered.add(60)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (45 < x < 70 and 45 < y < 70 or 25 < z < 45):
        triggered.add(61)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (45 < x < 35 and 45 < y < 70 and 25 < z < 45):
        triggered.add(62)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (20 < x < 70 and 45 < y < 70 and 25 < z < 45):
        triggered.add(63)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (45 < x < 70 and 45 < y < 55 and 25 < z < 45):
        triggered.add(64)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (45 < x < 70 and 45 < y < 70 and 15 < z < 45):
        triggered.add(65)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (45 < x < 70 and 45 < y < 70 and 25 < z < 35):
        triggered.add(66)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (45 < x < 60 and 45 < y < 70 and 25 < z < 45):
        triggered.add(67)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (15 < x < 70 and 45 < y < 70 and 25 < z < 45):
        triggered.add(68)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 78 or y < 42 and 20 < z < 40):
        triggered.add(69)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 78 and y < 42 or 20 < z < 40):
        triggered.add(70)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 48 and y < 42 and 20 < z < 40):
        triggered.add(71)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 78 and y < 22 and 20 < z < 40):
        triggered.add(72)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 78 and y < 42 and 10 < z < 40):
        triggered.add(73)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 78 and y < 42 and 20 < z < 30):
        triggered.add(74)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 78 and y < 42 and 20 < z < 50):
        triggered.add(75)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 78 and z < 42 and 20 < z < 40):
        triggered.add(76)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 78 and y < 42 and 20 < y < 40):
        triggered.add(77)
    if (x > 78 and y < 42 and 20 < z < 40) != (x + z > 98 and y < 42 and 20 < z < 40):
        triggered.add(78)
    if (x > y + 30) != (x > y + 10):
        triggered.add(79)
    if (x > y + 30) != (x > y + 20):
        triggered.add(80)
    if (x > y + 30) != (x > y + 40):
        triggered.add(81)
    if (x > y + 30) != (x > y):
        triggered.add(82)
    if (x > y + 30) != (x + 10 > y + 30):
        triggered.add(83)
    if (x > y + 30) != (x > z + 30):
        triggered.add(84)
    if (x > y + 30) != (z > y + 30):
        triggered.add(85)
    if (x > y + 30) != (x > y + z):
        triggered.add(86)
    if (x > y + 30) != (x > y - z):
        triggered.add(87)
    if (x > y + 30) != (x + z > y + 30):
        triggered.add(88)
    if (abs(x - y) < 10) != (abs(x + y) < 10):
        triggered.add(89)
    if (abs(x - y) < 10) != (abs(x - y) < 15):
        triggered.add(90)
    if (abs(x - y) < 10) != (abs(x - y) < 16):
        triggered.add(91)
    if (abs(x - y) < 10) != (abs(x - y) < 17):
        triggered.add(92)
    if (abs(x - y) < 10) != (abs(x - y) < 20):
        triggered.add(93)
    if (abs(x - y) < 10) != (abs(x - z) < 10):
        triggered.add(94)
    if (abs(x - z) < 15) != (abs(x + z) < 15):
        triggered.add(95)
    if (abs(x - z) < 15) != (abs(x - z) < 25):
        triggered.add(96)
    if (abs(x - z) < 15) != (abs(x - z) < 5):
        triggered.add(97)
    if (abs(x - z) < 15) != (abs(x - z) <= 15):
        triggered.add(98)
    if (abs(x - z) < 15) != (abs(x - y) < 15):
        triggered.add(99)
    if (abs(x - z) < 15) != (abs(y - z) < 15):
        triggered.add(100)
    if (45 < x < 70 and 45 < y < 70 and z < 12) != (45 < x < 70 or 45 < y < 70 and z < 12):
        triggered.add(101)
    if (45 < x < 70 and 45 < y < 70 and z < 12) != (45 < x < 70 and 45 < y < 70 or z < 12):
        triggered.add(102)
    if (45 < x < 70 and 45 < y < 70 and z < 12) != (25 < x < 70 and 45 < y < 70 and z < 12):
        triggered.add(103)
    if (45 < x < 70 and 45 < y < 70 and z < 12) != (45 < x < 50 and 45 < y < 70 and z < 12):
        triggered.add(104)
    if (45 < x < 70 and 45 < y < 70 and z < 12) != (45 < x < 70 and 25 < y < 70 and z < 12):
        triggered.add(105)
    if (45 < x < 70 and 45 < y < 70 and z < 12) != (45 < x < 70 and 45 < y < 50 and z < 12):
        triggered.add(106)
    if (45 < x < 70 and 45 < y < 70 and z < 12) != (45 < x < 70 and 45 < y < 70 and z < 10):
        triggered.add(107)
    if (45 < x < 70 and 45 < y < 70 and z < 12) != (45 < x < 70 and 45 < y < 70 and z < 20):
        triggered.add(108)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 70 or 45 < y < 70 and z > 55):
        triggered.add(109)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 70 and 45 < y < 70 or z > 55):
        triggered.add(110)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (25 < x < 70 and 45 < y < 70 and z > 55):
        triggered.add(111)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 60 and 45 < y < 70 and z > 55):
        triggered.add(112)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 70 and 40 < y < 70 and z > 55):
        triggered.add(113)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 70 and 45 < y < 80 and z > 55):
        triggered.add(114)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 70 and 45 < y < 70 and z > 45):
        triggered.add(115)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 70 and 45 < y < 70 and z > 35):
        triggered.add(116)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 70 and 45 < z < 70 and z > 55):
        triggered.add(117)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 70 and 45 < y < 70 and y > 55):
        triggered.add(118)
    if (x < 22 and y > 92 and 25 < z < 45) != (x < 22 or y > 92 and 25 < z < 45):
        triggered.add(119)
    if (x < 22 and y > 92 and 25 < z < 45) != (x < 22 and y > 92 or 25 < z < 45):
        triggered.add(120)
    if (x < 22 and y > 92 and 25 < z < 45) != (x < 12 and y > 92 and 25 < z < 45):
        triggered.add(121)
    if (x < 22 and y > 92 and 25 < z < 45) != (x < 22 and y > 52 and 25 < z < 45):
        triggered.add(122)
    if (x < 22 and y > 92 and 25 < z < 45) != (x < 22 and y > 92 and 15 < z < 45):
        triggered.add(123)
    if (x < 22 and y > 92 and 25 < z < 45) != (x < 22 and y > 92 and 25 < z < 35):
        triggered.add(124)

    return triggered


def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """"""
    generated_path = execute_validation_rules(particle[0], particle[1], particle[2])

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # : x:1-50, y:1-50, z:1-50
        self.bounds = bounds if bounds else [(1, 100), (1, 100), (1, 60)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """target pathsPSO"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(particles[i][0], particles[i][1], particles[i][2]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(gbest_particle[0], gbest_particle[1], gbest_particle[2]),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """Path PSO"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - Path ")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations")
    print(f"Path : {len(target_paths)}")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"Path {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "" if result['success'] else f"({result['best_fitness']:.3f})"
        print(f"{status} | {result['time']:.2f}s | iterations{result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f": {success_count}/{len(target_paths)} ({success_rate:.1f}%) | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs}")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations, {len(target_paths)}Path ")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"---  {run_idx}/{num_runs} ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f": {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs} runcompleted | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 1:
    ws1 = wb.active
    ws1.title = ""
    ws1.sheet_view.showGridLines = False

    headers = ["Run", "", "", "", "Average Iterations", "(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f" {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # 2: Path
    ws2 = wb.create_sheet(title="Path ")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Path ID", "", "", "", "Average Iterations", "Minimum Iterations", "Maximum Iterations"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"Path  {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # 3:
    ws3 = wb.create_sheet(title="")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Path ", "", "(x,y,z)", "", "Iterations", "Path "]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"{run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # 4: target paths
    ws4 = wb.create_sheet(title="target paths")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Path ID", "target paths", ""]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"Path  {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f" : {filename}")
    print(f"{'=' * 70}")
    print(f":")
    print(f"  1.        - {len(all_results)} run")
    print(f"  2. Path        - Path ")
    print(f"  3.    -  runPath ")
    print(f"  4. target paths       - target paths")
    print(f"{'=' * 70}\n")

    return filename


def main():


    target_paths = [
        [19, 24, 37, 47, 48, 51, 54, 55, 57, 58, 59, 61, 69, 70, 79, 82, 87, 88, 90, 91, 92, 93, 94, 95, 97, 100, 119,
         120],
        [19, 24, 37, 47, 48, 51, 54, 55, 57, 58, 59, 61, 69, 70, 79, 82, 87, 88, 90, 91, 92, 93, 95, 97, 100, 119, 120],
        [19, 24, 37, 47, 48, 51, 54, 55, 58, 59, 61, 69, 70, 79, 82, 87, 88, 91, 92, 93, 95, 97, 99, 100, 119, 120],
        [19, 24, 37, 47, 48, 51, 54, 55, 58, 59, 61, 69, 70, 79, 82, 87, 88, 90, 91, 92, 93, 96, 98, 99, 119, 120],
        [6, 14, 18, 31, 48, 60, 61, 82, 87, 88, 90, 91, 92, 93, 96, 99, 100, 101, 102, 109, 110, 115, 116, 118],
        [5, 6, 10, 30, 31, 34, 36, 37, 39, 43, 44, 46, 48, 61, 69, 79, 82, 84, 87, 88, 90, 91, 92, 93, 99, 120],
        [19, 24, 37, 47, 48, 51, 54, 55, 56, 57, 58, 59, 61, 69, 70, 82, 87, 88, 89, 94, 95, 97, 100, 119, 120],
        [19, 20, 24, 47, 48, 51, 54, 55, 58, 59, 69, 70, 79, 82, 87, 88, 90, 91, 92, 93, 94, 95, 97, 100, 119],
        [19, 37, 53, 62, 64, 66, 70, 87, 91, 92, 93, 94, 95, 97, 99, 100, 101, 102, 109, 110, 116, 118, 120],
        [5, 6, 18, 19, 31, 44, 47, 48, 60, 79, 80, 82, 83, 87, 88, 94, 95, 97, 99, 100, 101, 109, 110, 117],
        [5, 6, 7, 18, 31, 44, 48, 60, 79, 80, 82, 83, 87, 88, 94, 95, 97, 99, 100, 101, 109, 110, 113, 117],
        [5, 6, 18, 31, 44, 47, 48, 60, 79, 80, 82, 83, 85, 87, 88, 94, 95, 97, 99, 100, 101, 109, 110, 117],
        [19, 24, 37, 47, 48, 51, 52, 53, 61, 69, 70, 79, 80, 82, 83, 85, 87, 88, 94, 95, 97, 99, 100, 120],
        [6, 37, 48, 62, 64, 66, 67, 79, 82, 87, 88, 90, 91, 92, 93, 99, 101, 102, 109, 110, 116, 118, 120],
        [6, 18, 30, 31, 32, 35, 44, 48, 60, 87, 88, 91, 92, 93, 94, 95, 99, 100, 101, 109, 110, 114, 117],
        [6, 18, 30, 31, 32, 35, 42, 43, 44, 48, 60, 87, 88, 94, 95, 97, 99, 100, 101, 109, 110, 114, 117],
        [5, 30, 31, 34, 36, 37, 39, 43, 44, 46, 53, 61, 69, 70, 76, 79, 82, 84, 87, 88, 91, 92, 93, 120],
        [5, 10, 30, 31, 34, 41, 43, 44, 46, 53, 61, 69, 70, 76, 82, 84, 87, 88, 90, 91, 92, 93, 99, 120],
        [5, 10, 11, 19, 20, 25, 30, 31, 36, 37, 43, 44, 69, 79, 82, 84, 86, 87, 90, 91, 92, 93, 99, 102],
        [5, 6, 18, 30, 31, 33, 43, 44, 45, 48, 69, 79, 82, 87, 88, 90, 91, 92, 93, 99, 100, 109, 110],
        [5, 6, 10, 16, 30, 31, 33, 36, 43, 44, 45, 48, 69, 79, 80, 82, 83, 84, 87, 88, 100, 109, 110],
        [5, 30, 31, 34, 36, 37, 38, 43, 44, 46, 53, 61, 69, 70, 76, 84, 87, 88, 89, 94, 99, 119, 120],
        [6, 14, 18, 31, 44, 48, 60, 61, 79, 82, 87, 88, 90, 91, 92, 93, 94, 95, 101, 102, 112, 118],
        [37, 52, 53, 60, 61, 69, 70, 71, 79, 80, 82, 83, 87, 88, 95, 97, 99, 100, 101, 109, 120],
        [5, 6, 18, 19, 31, 44, 47, 48, 60, 81, 84, 86, 94, 95, 97, 99, 100, 101, 109, 110, 117],
        [11, 20, 25, 60, 61, 79, 82, 84, 86, 87, 90, 91, 92, 93, 99, 104, 106, 107, 109, 110]
    ]

    print("=" * 70)
    print("baseline PSO")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("Program completed")


if __name__ == "__main__":
    main()