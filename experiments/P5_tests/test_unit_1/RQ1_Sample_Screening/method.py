import os
import random
from collections import deque
import numpy as np
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"


def generate_input():
    return [
        random.randint(1, 100),
        random.randint(1, 100),
        random.randint(1, 60)
    ]


def execute_validation_rules(x, y, z):
    triggered = set()
    actions = []

    devices = {
        'main_light': 'green',
        'side_light': 'red',
        'pedestrian_light': 'red',
        'warning_system': 'off'
    }

    # Fixed syntax: properly formatted if statements
    if (x > 85 and y < 40 and z < 25) != (x > 90 and y < 40 and z < 25):
        triggered.add(1)
    if (x > 85 and y < 40 and z < 25) != (x > 85 and y < 35 and z < 25):
        triggered.add(2)
    if (x > 85 and y < 40 and z < 25) != (x > 85 and y < 40 and z < 20):
        triggered.add(3)
    if (x > 85 and y < 40 and z < 25) != (x > 80 and y < 40 and z < 25):
        triggered.add(4)
    if (x > 80 and y < 45 and z > 40) != (x > 80 or y < 45 and z > 40):
        triggered.add(5)
    if (x > 80 and y < 45 and z > 40) != (x > 80 and y < 45 or z > 40):
        triggered.add(6)
    if (x > 80 and y < 45 and z > 40) != (x > 60 and y < 45 and z > 40):
        triggered.add(7)
    if (x > 80 and y < 45 and z > 40) != (x > 80 and y < 25 and z > 40):
        triggered.add(8)
    if (x > 80 and y < 45 and z > 40) != (x > 80 and y < 45 and z > 20):
        triggered.add(9)
    if (x > 92 and y < 30 and z < 15) != (x > 92 or y < 30 and z < 15):
        triggered.add(10)
    if (x > 92 and y < 30 and z < 15) != (x > 92 and y < 30 or z < 15):
        triggered.add(11)
    if (x > 92 and y < 30 and z < 15) != (x > 72 and y < 30 and z < 15):
        triggered.add(12)
    if (x > 92 and y < 30 and z < 15) != (x > 92 and y < 10 and z < 15):
        triggered.add(13)
    if (x > 75 and 45 < y < 65 and z > 50) != (x > 55 and 45 < y < 65 and z > 50):
        triggered.add(14)
    if (x > 75 and 45 < y < 65 and z > 50) != (x > 75 and 35 < y < 65 and z > 50):
        triggered.add(15)
    if (x > 75 and 45 < y < 65 and z > 50) != (x > 75 and 45 < y < 55 and z > 50):
        triggered.add(16)
    if (x > 75 and 45 < y < 65 and z > 50) != (x > 75 and 45 < y < 65 and z > 40):
        triggered.add(17)
    if (x > 75 and 45 < y < 65 and z > 50) != (x > 75 and 45 < y < 65 or z > 50):
        triggered.add(18)
    if (x < 50 and y > 80 and z < 25) != (x < 50 or y > 80 and z < 25):
        triggered.add(19)
    if (x < 50 and y > 80 and z < 25) != (x < 50 and y > 80 or z < 25):
        triggered.add(20)
    if (x < 50 and y > 80 and z < 25) != (x < 25 and y > 80 and z < 25):
        triggered.add(21)
    if (x < 50 and y > 80 and z < 25) != (x < 50 and y > 40 and z < 25):
        triggered.add(22)
    if (x < 50 and y > 80 and z < 25) != (x < 50 and y > 80 and z < 15):
        triggered.add(23)
    if (x < 30 and y > 92 and z < 15) != (x < 30 or y > 92 and z < 15):
        triggered.add(24)
    if (x < 30 and y > 92 and z < 15) != (x < 30 and y > 92 or z < 15):
        triggered.add(25)
    if (x < 30 and y > 92 and z < 15) != (x < 70 and y > 92 and z < 15):
        triggered.add(26)
    if (x < 30 and y > 92 and z < 15) != (x < 30 and y > 42 and z < 15):
        triggered.add(27)
    if (x < 30 and y > 92 and z < 15) != (x < 30 and y > 92 and z < 5):
        triggered.add(28)
    if (x < 30 and y > 92 and z < 15) != (x < 30 and y > 62 and z < 15):
        triggered.add(29)
    if (x > 70 and y > 70 and z > 45) != (x > 70 or y > 70 and z > 45):
        triggered.add(30)
    if (x > 70 and y > 70 and z > 45) != (x > 70 and y > 70 or z > 45):
        triggered.add(31)
    if (x > 70 and y > 70 and z > 45) != (x > 50 and y > 70 and z > 45):
        triggered.add(32)
    if (x > 70 and y > 70 and z > 45) != (x > 70 and y > 50 and z > 45):
        triggered.add(33)
    if (x > 70 and y > 70 and z > 45) != (x > 70 and y > 70 and z > 25):
        triggered.add(34)
    if (x > 70 and y > 70 and z > 45) != (x > 35 and y > 70 and z > 45):
        triggered.add(35)
    if (x > 88 and y > 88 and 25 < z < 45) != (x > 88 or y > 88 and 25 < z < 45):
        triggered.add(36)
    if (x > 88 and y > 88 and 25 < z < 45) != (x > 88 and y > 88 or 25 < z < 45):
        triggered.add(37)
    if (x > 88 and y > 88 and 25 < z < 45) != (x > 44 and y > 88 and 25 < z < 45):
        triggered.add(38)
    if (x > 88 and y > 88 and 25 < z < 45) != (x > 88 and y > 44 and 25 < z < 45):
        triggered.add(39)
    if (x > 88 and y > 88 and 25 < z < 45) != (x > 88 and y > 88 and 15 < z < 45):
        triggered.add(40)
    if (x > 88 and y > 88 and 25 < z < 45) != (x > 88 and y > 88 and 25 < z < 25):
        triggered.add(41)
    if (x > 75 and y > 75 and z > 55) != (x > 35 and y > 75 and z > 55):
        triggered.add(42)
    if (x > 75 and y > 75 and z > 55) != (x > 75 or y > 75 and z > 55):
        triggered.add(43)
    if (x > 75 and y > 75 and z > 55) != (x > 75 and y > 75 or z > 55):
        triggered.add(44)
    if (x > 75 and y > 75 and z > 55) != (x > 75 and y > 35 and z > 55):
        triggered.add(45)
    if (x > 75 and y > 75 and z > 55) != (x > 75 and y > 75 and z > 25):
        triggered.add(46)
    if (x < 40 and y < 40 and z > 40) != (x < 40 or y < 40 and z > 40):
        triggered.add(47)
    if (x < 40 and y < 40 and z > 40) != (x < 40 and y < 40 or z > 40):
        triggered.add(48)
    if (x < 40 and y < 40 and z > 40) != (x < 20 and y < 40 and z > 40):
        triggered.add(49)
    if (x < 40 and y < 40 and z > 40) != (x < 40 and y < 20 and z > 40):
        triggered.add(50)
    if (x < 40 and y < 40 and z > 40) != (x < 40 and y < 40 and z > 20):
        triggered.add(51)
    if (x < 25 and y < 25 and 20 < z < 40) != (x < 25 or y < 25 and 20 < z < 40):
        triggered.add(52)
    if (x < 25 and y < 25 and 20 < z < 40) != (x < 25 and y < 25 or 20 < z < 40):
        triggered.add(53)
    if (x < 25 and y < 25 and 20 < z < 40) != (x < 15 and y < 25 and 20 < z < 40):
        triggered.add(54)
    if (x < 25 and y < 25 and 20 < z < 40) != (x < 5 and y < 25 and 20 < z < 40):
        triggered.add(55)
    if (x < 25 and y < 25 and 20 < z < 40) != (x < 25 and y < 15 and 20 < z < 40):
        triggered.add(56)
    if (x < 25 and y < 25 and 20 < z < 40) != (x < 25 and y < 5 and 20 < z < 40):
        triggered.add(57)
    if (x < 25 and y < 25 and 20 < z < 40) != (x < 25 and y < 25 and 20 < z < 20):
        triggered.add(58)
    if (x < 25 and y < 25 and 20 < z < 40) != (x < 25 and y < 25 and 20 < z < 10):
        triggered.add(59)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (45 < x < 70 or 45 < y < 70 and 25 < z < 45):
        triggered.add(60)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (45 < x < 70 and 45 < y < 70 or 25 < z < 45):
        triggered.add(61)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (45 < x < 35 and 45 < y < 70 and 25 < z < 45):
        triggered.add(62)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (20 < x < 70 and 45 < y < 70 and 25 < z < 45):
        triggered.add(63)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (45 < x < 70 and 45 < y < 55 and 25 < z < 45):
        triggered.add(64)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (45 < x < 70 and 45 < y < 70 and 15 < z < 45):
        triggered.add(65)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (45 < x < 70 and 45 < y < 70 and 25 < z < 35):
        triggered.add(66)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (45 < x < 60 and 45 < y < 70 and 25 < z < 45):
        triggered.add(67)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (15 < x < 70 and 45 < y < 70 and 25 < z < 45):
        triggered.add(68)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 78 or y < 42 and 20 < z < 40):
        triggered.add(69)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 78 and y < 42 or 20 < z < 40):
        triggered.add(70)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 48 and y < 42 and 20 < z < 40):
        triggered.add(71)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 78 and y < 22 and 20 < z < 40):
        triggered.add(72)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 78 and y < 42 and 10 < z < 40):
        triggered.add(73)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 78 and y < 42 and 20 < z < 30):
        triggered.add(74)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 78 and y < 42 and 20 < z < 50):
        triggered.add(75)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 78 and z < 42 and 20 < z < 40):
        triggered.add(76)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 78 and y < 42 and 20 < y < 40):
        triggered.add(77)
    if (x > 78 and y < 42 and 20 < z < 40) != (x + z > 98 and y < 42 and 20 < z < 40):
        triggered.add(78)
    if (x > y + 30) != (x > y + 10):
        triggered.add(79)
    if (x > y + 30) != (x > y + 20):
        triggered.add(80)
    if (x > y + 30) != (x > y + 40):
        triggered.add(81)
    if (x > y + 30) != (x > y):
        triggered.add(82)
    if (x > y + 30) != (x + 10 > y + 30):
        triggered.add(83)
    if (x > y + 30) != (x > z + 30):
        triggered.add(84)
    if (x > y + 30) != (z > y + 30):
        triggered.add(85)
    if (x > y + 30) != (x > y + z):
        triggered.add(86)
    if (x > y + 30) != (x > y - z):
        triggered.add(87)
    if (x > y + 30) != (x + z > y + 30):
        triggered.add(88)
    if (abs(x - y) < 10) != (abs(x + y) < 10):
        triggered.add(89)
    if (abs(x - y) < 10) != (abs(x - y) < 15):
        triggered.add(90)
    if (abs(x - y) < 10) != (abs(x - y) < 16):
        triggered.add(91)
    if (abs(x - y) < 10) != (abs(x - y) < 17):
        triggered.add(92)
    if (abs(x - y) < 10) != (abs(x - y) < 20):
        triggered.add(93)
    if (abs(x - y) < 10) != (abs(x - z) < 10):
        triggered.add(94)
    if (abs(x - z) < 15) != (abs(x + z) < 15):
        triggered.add(95)
    if (abs(x - z) < 15) != (abs(x - z) < 25):
        triggered.add(96)
    if (abs(x - z) < 15) != (abs(x - z) < 5):
        triggered.add(97)
    if (abs(x - z) < 15) != (abs(x - z) <= 15):
        triggered.add(98)
    if (abs(x - z) < 15) != (abs(x - y) < 15):
        triggered.add(99)
    if (abs(x - z) < 15) != (abs(y - z) < 15):
        triggered.add(100)
    if (45 < x < 70 and 45 < y < 70 and z < 12) != (45 < x < 70 or 45 < y < 70 and z < 12):
        triggered.add(101)
    if (45 < x < 70 and 45 < y < 70 and z < 12) != (45 < x < 70 and 45 < y < 70 or z < 12):
        triggered.add(102)
    if (45 < x < 70 and 45 < y < 70 and z < 12) != (25 < x < 70 and 45 < y < 70 and z < 12):
        triggered.add(103)
    if (45 < x < 70 and 45 < y < 70 and z < 12) != (45 < x < 50 and 45 < y < 70 and z < 12):
        triggered.add(104)
    if (45 < x < 70 and 45 < y < 70 and z < 12) != (45 < x < 70 and 25 < y < 70 and z < 12):
        triggered.add(105)
    if (45 < x < 70 and 45 < y < 70 and z < 12) != (45 < x < 70 and 45 < y < 50 and z < 12):
        triggered.add(106)
    if (45 < x < 70 and 45 < y < 70 and z < 12) != (45 < x < 70 and 45 < y < 70 and z < 10):
        triggered.add(107)
    if (45 < x < 70 and 45 < y < 70 and z < 12) != (45 < x < 70 and 45 < y < 70 and z < 20):
        triggered.add(108)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 70 or 45 < y < 70 and z > 55):
        triggered.add(109)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 70 and 45 < y < 70 or z > 55):
        triggered.add(110)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (25 < x < 70 and 45 < y < 70 and z > 55):
        triggered.add(111)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 60 and 45 < y < 70 and z > 55):
        triggered.add(112)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 70 and 40 < y < 70 and z > 55):
        triggered.add(113)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 70 and 45 < y < 80 and z > 55):
        triggered.add(114)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 70 and 45 < y < 70 and z > 45):
        triggered.add(115)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 70 and 45 < y < 70 and z > 35):
        triggered.add(116)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 70 and 45 < z < 70 and z > 55):
        triggered.add(117)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 70 and 45 < y < 70 and y > 55):
        triggered.add(118)
    if (x < 22 and y > 92 and 25 < z < 45) != (x < 22 or y > 92 and 25 < z < 45):
        triggered.add(119)
    if (x < 22 and y > 92 and 25 < z < 45) != (x < 22 and y > 92 or 25 < z < 45):
        triggered.add(120)
    if (x < 22 and y > 92 and 25 < z < 45) != (x < 12 and y > 92 and 25 < z < 45):
        triggered.add(121)
    if (x < 22 and y > 92 and 25 < z < 45) != (x < 22 and y > 52 and 25 < z < 45):
        triggered.add(122)
    if (x < 22 and y > 92 and 25 < z < 45) != (x < 22 and y > 92 and 15 < z < 45):
        triggered.add(123)
    if (x < 22 and y > 92 and 25 < z < 45) != (x < 22 and y > 92 and 25 < z < 35):
        triggered.add(124)

    return triggered


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0


# Fixed: Changed variable name from targetPaths to TARGET_PATHS and fixed list syntax
TARGET_PATHS = [
    [19, 24, 37, 47, 48, 51, 54, 55, 57, 58, 59, 61, 69, 70, 79, 82, 87, 88, 90, 91, 92, 93, 94, 95, 97, 100, 119, 120],
    [19, 24, 37, 47, 48, 51, 54, 55, 57, 58, 59, 61, 69, 70, 79, 82, 87, 88, 90, 91, 92, 93, 95, 97, 100, 119, 120],
    [19, 24, 37, 47, 48, 51, 54, 55, 58, 59, 61, 69, 70, 79, 82, 87, 88, 91, 92, 93, 95, 97, 99, 100, 119, 120],
    [19, 24, 37, 47, 48, 51, 54, 55, 58, 59, 61, 69, 70, 79, 82, 87, 88, 90, 91, 92, 93, 96, 98, 99, 119, 120],
    [6, 14, 18, 31, 48, 60, 61, 82, 87, 88, 90, 91, 92, 93, 96, 99, 100, 101, 102, 109, 110, 115, 116, 118],
    [5, 6, 10, 30, 31, 34, 36, 37, 39, 43, 44, 46, 48, 61, 69, 79, 82, 84, 87, 88, 90, 91, 92, 93, 99, 120],
    [19, 24, 37, 47, 48, 51, 54, 55, 56, 57, 58, 59, 61, 69, 70, 82, 87, 88, 89, 94, 95, 97, 100, 119, 120],
    [19, 20, 24, 47, 48, 51, 54, 55, 58, 59, 69, 70, 79, 82, 87, 88, 90, 91, 92, 93, 94, 95, 97, 100, 119],
    [19, 37, 53, 62, 64, 66, 70, 87, 91, 92, 93, 94, 95, 97, 99, 100, 101, 102, 109, 110, 116, 118, 120],
    [5, 6, 18, 19, 31, 44, 47, 48, 60, 79, 80, 82, 83, 87, 88, 94, 95, 97, 99, 100, 101, 109, 110, 117],
    [5, 6, 7, 18, 31, 44, 48, 60, 79, 80, 82, 83, 87, 88, 94, 95, 97, 99, 100, 101, 109, 110, 113, 117],
    [5, 6, 18, 31, 44, 47, 48, 60, 79, 80, 82, 83, 85, 87, 88, 94, 95, 97, 99, 100, 101, 109, 110, 117],
    [19, 24, 37, 47, 48, 51, 52, 53, 61, 69, 70, 79, 80, 82, 83, 85, 87, 88, 94, 95, 97, 99, 100, 120],
    [6, 37, 48, 62, 64, 66, 67, 79, 82, 87, 88, 90, 91, 92, 93, 99, 101, 102, 109, 110, 116, 118, 120],
    [6, 18, 30, 31, 32, 35, 44, 48, 60, 87, 88, 91, 92, 93, 94, 95, 99, 100, 101, 109, 110, 114, 117],
    [6, 18, 30, 31, 32, 35, 42, 43, 44, 48, 60, 87, 88, 94, 95, 97, 99, 100, 101, 109, 110, 114, 117],
    [5, 30, 31, 34, 36, 37, 39, 43, 44, 46, 53, 61, 69, 70, 76, 79, 82, 84, 87, 88, 91, 92, 93, 120],
    [5, 10, 30, 31, 34, 41, 43, 44, 46, 53, 61, 69, 70, 76, 82, 84, 87, 88, 90, 91, 92, 93, 99, 120],
    [5, 10, 11, 19, 20, 25, 30, 31, 36, 37, 43, 44, 69, 79, 82, 84, 86, 87, 90, 91, 92, 93, 99, 102],
    [5, 6, 18, 30, 31, 33, 43, 44, 45, 48, 69, 79, 82, 87, 88, 90, 91, 92, 93, 99, 100, 109, 110],
    [5, 6, 10, 16, 30, 31, 33, 36, 43, 44, 45, 48, 69, 79, 80, 82, 83, 84, 87, 88, 100, 109, 110],
    [5, 30, 31, 34, 36, 37, 38, 43, 44, 46, 53, 61, 69, 70, 76, 84, 87, 88, 89, 94, 99, 119, 120],
    [6, 14, 18, 31, 44, 48, 60, 61, 79, 82, 87, 88, 90, 91, 92, 93, 94, 95, 101, 102, 112, 118],
    [37, 52, 53, 60, 61, 69, 70, 71, 79, 80, 82, 83, 87, 88, 95, 97, 99, 100, 101, 109, 120],
    [5, 6, 18, 19, 31, 44, 47, 48, 60, 81, 84, 86, 94, 95, 97, 99, 100, 101, 109, 110, 117],
    [11, 20, 25, 60, 61, 79, 82, 84, 86, 87, 90, 91, 92, 93, 99, 104, 106, 107, 109, 110]
]


class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 2000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(TARGET_PATHS)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }


def compute_robustness(state, path):
    # Fixed: Changed execute_Tr to execute_validation_rules
    base = execute_validation_rules(state[0], state[1], state[2])
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                    np.clip(state[0] + dx, 1, 30),
                    np.clip(state[1] + dy, 1, 40),
                    np.clip(state[2] + dz, 1, 2100)
                ])
                n_trig = execute_validation_rules(neighbor[0], neighbor[1], neighbor[2])
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = TARGET_PATHS[target_path_idx]
    samples = []
    attempts = 0

    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(1, 30),
            random.randint(1, 40),
            random.randint(1, 2100)
        ])
        triggered = execute_validation_rules(state[0], state[1], state[2])

        if not triggered:
            continue

        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)

        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }

        samples.append(sample_data)

    return samples


def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample[
                'robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)

        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]

    return selected_samples


def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = TARGET_PATHS[target_path_idx]
        samples = []
        attempts = 0

        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(1, 30),
                random.randint(1, 40),
                random.randint(1, 2100)
            ])
            triggered = execute_validation_rules(state[0], state[1], state[2])

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)

            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)

        random.shuffle(samples)
        return samples[:config.top_k_samples]

    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")

        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)


def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}

    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates

    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}

        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config,
                                                         shared_candidates[path_idx])

            strategy_results[path_idx] = samples

        results[strategy_name] = strategy_results

    return results


def analyze_fitness_values(results, config):
    analysis_results = {}

    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []

        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])

        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }

        all_scores = []

        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])

        scores_array = np.array(all_scores)

        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)

        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })

        analysis_results[strategy_name] = analysis

    return analysis_results


def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []

    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)

    df = pd.DataFrame(df_data)
    return df, analysis_results


def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []

    print(f"Starting {num_runs} experiments...")

    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")

        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)

        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)

        print(f"Experiment {run_idx + 1} completed.")

    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df


def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')

    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width

    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []

    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]

        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)

    stats_df = pd.DataFrame(stats_data)

    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Results saved to: {output_path}")


def main():
    results_df = run_multiple_experiments(num_runs=20)

    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")

    save_results_to_excel(results_df, output_path)

    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)

    return results_df, output_path


if __name__ == "__main__":
    results, output_file = main()