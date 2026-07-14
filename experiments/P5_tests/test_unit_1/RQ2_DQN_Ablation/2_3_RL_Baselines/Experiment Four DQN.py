
import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
import numpy as np
import random
import time
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# ===  ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'MIN_X': 1, 'MAX_X': 100,
    'MIN_Y': 1, 'MAX_Y': 100,
    'MIN_Z': 1, 'MAX_Z': 100,
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 5,
    'NUM_ROUNDS': 5,
    'REPLAY_BATCH_SIZE': 64,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'LEARNING_RATE': 3e-4,
    'NUM_RUNS': 20,  # 20 run
    'TOP_K_SAMPLES': 20,
    'REPLAY_BUFFER_CAPACITY': 20000,  # Path
    'TARGET_PATHS': [
        [19, 24, 37, 47, 48, 51, 54, 55, 57, 58, 59, 61, 69, 70, 79, 82, 87, 88, 90, 91, 92, 93, 94, 95, 97, 100, 119,
         120],
        [19, 24, 37, 47, 48, 51, 54, 55, 57, 58, 59, 61, 69, 70, 79, 82, 87, 88, 90, 91, 92, 93, 95, 97, 100, 119, 120],
        [19, 24, 37, 47, 48, 51, 54, 55, 58, 59, 61, 69, 70, 79, 82, 87, 88, 91, 92, 93, 95, 97, 99, 100, 119, 120],
        [19, 24, 37, 47, 48, 51, 54, 55, 58, 59, 61, 69, 70, 79, 82, 87, 88, 90, 91, 92, 93, 96, 98, 99, 119, 120],
        [6, 14, 18, 31, 48, 60, 61, 82, 87, 88, 90, 91, 92, 93, 96, 99, 100, 101, 102, 109, 110, 115, 116, 118],
        [5, 6, 10, 30, 31, 34, 36, 37, 39, 43, 44, 46, 48, 61, 69, 79, 82, 84, 87, 88, 90, 91, 92, 93, 99, 120],
        [19, 24, 37, 47, 48, 51, 54, 55, 56, 57, 58, 59, 61, 69, 70, 82, 87, 88, 89, 94, 95, 97, 100, 119, 120],
        [19, 20, 24, 47, 48, 51, 54, 55, 58, 59, 69, 70, 79, 82, 87, 88, 90, 91, 92, 93, 94, 95, 97, 100, 119],
        [19, 37, 53, 62, 64, 66, 70, 87, 91, 92, 93, 94, 95, 97, 99, 100, 101, 102, 109, 110, 116, 118, 120],
        [5, 6, 18, 19, 31, 44, 47, 48, 60, 79, 80, 82, 83, 87, 88, 94, 95, 97, 99, 100, 101, 109, 110, 117],
        [5, 6, 7, 18, 31, 44, 48, 60, 79, 80, 82, 83, 87, 88, 94, 95, 97, 99, 100, 101, 109, 110, 113, 117],
        [5, 6, 18, 31, 44, 47, 48, 60, 79, 80, 82, 83, 85, 87, 88, 94, 95, 97, 99, 100, 101, 109, 110, 117],
        [19, 24, 37, 47, 48, 51, 52, 53, 61, 69, 70, 79, 80, 82, 83, 85, 87, 88, 94, 95, 97, 99, 100, 120],
        [6, 37, 48, 62, 64, 66, 67, 79, 82, 87, 88, 90, 91, 92, 93, 99, 101, 102, 109, 110, 116, 118, 120],
        [6, 18, 30, 31, 32, 35, 44, 48, 60, 87, 88, 91, 92, 93, 94, 95, 99, 100, 101, 109, 110, 114, 117],
        [6, 18, 30, 31, 32, 35, 42, 43, 44, 48, 60, 87, 88, 94, 95, 97, 99, 100, 101, 109, 110, 114, 117],
        [5, 30, 31, 34, 36, 37, 39, 43, 44, 46, 53, 61, 69, 70, 76, 79, 82, 84, 87, 88, 91, 92, 93, 120],
        [5, 10, 30, 31, 34, 41, 43, 44, 46, 53, 61, 69, 70, 76, 82, 84, 87, 88, 90, 91, 92, 93, 99, 120],
        [5, 10, 11, 19, 20, 25, 30, 31, 36, 37, 43, 44, 69, 79, 82, 84, 86, 87, 90, 91, 92, 93, 99, 102],
        [5, 6, 18, 30, 31, 33, 43, 44, 45, 48, 69, 79, 82, 87, 88, 90, 91, 92, 93, 99, 100, 109, 110],
        [5, 6, 10, 16, 30, 31, 33, 36, 43, 44, 45, 48, 69, 79, 80, 82, 83, 84, 87, 88, 100, 109, 110],
        [5, 30, 31, 34, 36, 37, 38, 43, 44, 46, 53, 61, 69, 70, 76, 84, 87, 88, 89, 94, 99, 119, 120],
        [6, 14, 18, 31, 44, 48, 60, 61, 79, 82, 87, 88, 90, 91, 92, 93, 94, 95, 101, 102, 112, 118],
        [37, 52, 53, 60, 61, 69, 70, 71, 79, 80, 82, 83, 87, 88, 95, 97, 99, 100, 101, 109, 120],
        [5, 6, 18, 19, 31, 44, 47, 48, 60, 81, 84, 86, 94, 95, 97, 99, 100, 101, 109, 110, 117],
        [11, 20, 25, 60, 61, 79, 82, 84, 86, 87, 90, 91, 92, 93, 99, 104, 106, 107, 109, 110]
    ],
}


# ===  ===
def get_bounds():
    mins = np.array([EXPERIMENT_CONFIG['MIN_X'], EXPERIMENT_CONFIG['MIN_Y'], EXPERIMENT_CONFIG['MIN_Z']])
    maxs = np.array([EXPERIMENT_CONFIG['MAX_X'], EXPERIMENT_CONFIG['MAX_Y'], EXPERIMENT_CONFIG['MAX_Z']])
    return mins, maxs

def clip_state(state):
    mins, maxs = get_bounds()
    return np.clip(state, mins, maxs)

def denormalize_state(normalized_state):
    """"""
    mins, maxs = get_bounds()
    return normalized_state * (maxs - mins) / 2 + (mins + maxs) / 2


def coverage_similarity(triggered, target_path):
    """
    Similarity: / target paths
    """
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0

    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)


def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']

    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']

    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']

    return reward


# ===   ===
def execute_Tr(x, y, z):
    triggered = set()
    actions = []

    devices = {
        'main_light': 'green',
        'side_light': 'red',
        'pedestrian_light': 'red',
        'warning_system': 'off'
    }

    # Fixed syntax: properly formatted if statements
    if (x > 85 and y < 40 and z < 25) != (x > 90 and y < 40 and z < 25):
        triggered.add(1)
    if (x > 85 and y < 40 and z < 25) != (x > 85 and y < 35 and z < 25):
        triggered.add(2)
    if (x > 85 and y < 40 and z < 25) != (x > 85 and y < 40 and z < 20):
        triggered.add(3)
    if (x > 85 and y < 40 and z < 25) != (x > 80 and y < 40 and z < 25):
        triggered.add(4)
    if (x > 80 and y < 45 and z > 40) != (x > 80 or y < 45 and z > 40):
        triggered.add(5)
    if (x > 80 and y < 45 and z > 40) != (x > 80 and y < 45 or z > 40):
        triggered.add(6)
    if (x > 80 and y < 45 and z > 40) != (x > 60 and y < 45 and z > 40):
        triggered.add(7)
    if (x > 80 and y < 45 and z > 40) != (x > 80 and y < 25 and z > 40):
        triggered.add(8)
    if (x > 80 and y < 45 and z > 40) != (x > 80 and y < 45 and z > 20):
        triggered.add(9)
    if (x > 92 and y < 30 and z < 15) != (x > 92 or y < 30 and z < 15):
        triggered.add(10)
    if (x > 92 and y < 30 and z < 15) != (x > 92 and y < 30 or z < 15):
        triggered.add(11)
    if (x > 92 and y < 30 and z < 15) != (x > 72 and y < 30 and z < 15):
        triggered.add(12)
    if (x > 92 and y < 30 and z < 15) != (x > 92 and y < 10 and z < 15):
        triggered.add(13)
    if (x > 75 and 45 < y < 65 and z > 50) != (x > 55 and 45 < y < 65 and z > 50):
        triggered.add(14)
    if (x > 75 and 45 < y < 65 and z > 50) != (x > 75 and 35 < y < 65 and z > 50):
        triggered.add(15)
    if (x > 75 and 45 < y < 65 and z > 50) != (x > 75 and 45 < y < 55 and z > 50):
        triggered.add(16)
    if (x > 75 and 45 < y < 65 and z > 50) != (x > 75 and 45 < y < 65 and z > 40):
        triggered.add(17)
    if (x > 75 and 45 < y < 65 and z > 50) != (x > 75 and 45 < y < 65 or z > 50):
        triggered.add(18)
    if (x < 50 and y > 80 and z < 25) != (x < 50 or y > 80 and z < 25):
        triggered.add(19)
    if (x < 50 and y > 80 and z < 25) != (x < 50 and y > 80 or z < 25):
        triggered.add(20)
    if (x < 50 and y > 80 and z < 25) != (x < 25 and y > 80 and z < 25):
        triggered.add(21)
    if (x < 50 and y > 80 and z < 25) != (x < 50 and y > 40 and z < 25):
        triggered.add(22)
    if (x < 50 and y > 80 and z < 25) != (x < 50 and y > 80 and z < 15):
        triggered.add(23)
    if (x < 30 and y > 92 and z < 15) != (x < 30 or y > 92 and z < 15):
        triggered.add(24)
    if (x < 30 and y > 92 and z < 15) != (x < 30 and y > 92 or z < 15):
        triggered.add(25)
    if (x < 30 and y > 92 and z < 15) != (x < 70 and y > 92 and z < 15):
        triggered.add(26)
    if (x < 30 and y > 92 and z < 15) != (x < 30 and y > 42 and z < 15):
        triggered.add(27)
    if (x < 30 and y > 92 and z < 15) != (x < 30 and y > 92 and z < 5):
        triggered.add(28)
    if (x < 30 and y > 92 and z < 15) != (x < 30 and y > 62 and z < 15):
        triggered.add(29)
    if (x > 70 and y > 70 and z > 45) != (x > 70 or y > 70 and z > 45):
        triggered.add(30)
    if (x > 70 and y > 70 and z > 45) != (x > 70 and y > 70 or z > 45):
        triggered.add(31)
    if (x > 70 and y > 70 and z > 45) != (x > 50 and y > 70 and z > 45):
        triggered.add(32)
    if (x > 70 and y > 70 and z > 45) != (x > 70 and y > 50 and z > 45):
        triggered.add(33)
    if (x > 70 and y > 70 and z > 45) != (x > 70 and y > 70 and z > 25):
        triggered.add(34)
    if (x > 70 and y > 70 and z > 45) != (x > 35 and y > 70 and z > 45):
        triggered.add(35)
    if (x > 88 and y > 88 and 25 < z < 45) != (x > 88 or y > 88 and 25 < z < 45):
        triggered.add(36)
    if (x > 88 and y > 88 and 25 < z < 45) != (x > 88 and y > 88 or 25 < z < 45):
        triggered.add(37)
    if (x > 88 and y > 88 and 25 < z < 45) != (x > 44 and y > 88 and 25 < z < 45):
        triggered.add(38)
    if (x > 88 and y > 88 and 25 < z < 45) != (x > 88 and y > 44 and 25 < z < 45):
        triggered.add(39)
    if (x > 88 and y > 88 and 25 < z < 45) != (x > 88 and y > 88 and 15 < z < 45):
        triggered.add(40)
    if (x > 88 and y > 88 and 25 < z < 45) != (x > 88 and y > 88 and 25 < z < 25):
        triggered.add(41)
    if (x > 75 and y > 75 and z > 55) != (x > 35 and y > 75 and z > 55):
        triggered.add(42)
    if (x > 75 and y > 75 and z > 55) != (x > 75 or y > 75 and z > 55):
        triggered.add(43)
    if (x > 75 and y > 75 and z > 55) != (x > 75 and y > 75 or z > 55):
        triggered.add(44)
    if (x > 75 and y > 75 and z > 55) != (x > 75 and y > 35 and z > 55):
        triggered.add(45)
    if (x > 75 and y > 75 and z > 55) != (x > 75 and y > 75 and z > 25):
        triggered.add(46)
    if (x < 40 and y < 40 and z > 40) != (x < 40 or y < 40 and z > 40):
        triggered.add(47)
    if (x < 40 and y < 40 and z > 40) != (x < 40 and y < 40 or z > 40):
        triggered.add(48)
    if (x < 40 and y < 40 and z > 40) != (x < 20 and y < 40 and z > 40):
        triggered.add(49)
    if (x < 40 and y < 40 and z > 40) != (x < 40 and y < 20 and z > 40):
        triggered.add(50)
    if (x < 40 and y < 40 and z > 40) != (x < 40 and y < 40 and z > 20):
        triggered.add(51)
    if (x < 25 and y < 25 and 20 < z < 40) != (x < 25 or y < 25 and 20 < z < 40):
        triggered.add(52)
    if (x < 25 and y < 25 and 20 < z < 40) != (x < 25 and y < 25 or 20 < z < 40):
        triggered.add(53)
    if (x < 25 and y < 25 and 20 < z < 40) != (x < 15 and y < 25 and 20 < z < 40):
        triggered.add(54)
    if (x < 25 and y < 25 and 20 < z < 40) != (x < 5 and y < 25 and 20 < z < 40):
        triggered.add(55)
    if (x < 25 and y < 25 and 20 < z < 40) != (x < 25 and y < 15 and 20 < z < 40):
        triggered.add(56)
    if (x < 25 and y < 25 and 20 < z < 40) != (x < 25 and y < 5 and 20 < z < 40):
        triggered.add(57)
    if (x < 25 and y < 25 and 20 < z < 40) != (x < 25 and y < 25 and 20 < z < 20):
        triggered.add(58)
    if (x < 25 and y < 25 and 20 < z < 40) != (x < 25 and y < 25 and 20 < z < 10):
        triggered.add(59)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (45 < x < 70 or 45 < y < 70 and 25 < z < 45):
        triggered.add(60)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (45 < x < 70 and 45 < y < 70 or 25 < z < 45):
        triggered.add(61)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (45 < x < 35 and 45 < y < 70 and 25 < z < 45):
        triggered.add(62)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (20 < x < 70 and 45 < y < 70 and 25 < z < 45):
        triggered.add(63)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (45 < x < 70 and 45 < y < 55 and 25 < z < 45):
        triggered.add(64)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (45 < x < 70 and 45 < y < 70 and 15 < z < 45):
        triggered.add(65)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (45 < x < 70 and 45 < y < 70 and 25 < z < 35):
        triggered.add(66)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (45 < x < 60 and 45 < y < 70 and 25 < z < 45):
        triggered.add(67)
    if (45 < x < 70 and 45 < y < 70 and 25 < z < 45) != (15 < x < 70 and 45 < y < 70 and 25 < z < 45):
        triggered.add(68)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 78 or y < 42 and 20 < z < 40):
        triggered.add(69)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 78 and y < 42 or 20 < z < 40):
        triggered.add(70)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 48 and y < 42 and 20 < z < 40):
        triggered.add(71)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 78 and y < 22 and 20 < z < 40):
        triggered.add(72)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 78 and y < 42 and 10 < z < 40):
        triggered.add(73)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 78 and y < 42 and 20 < z < 30):
        triggered.add(74)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 78 and y < 42 and 20 < z < 50):
        triggered.add(75)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 78 and z < 42 and 20 < z < 40):
        triggered.add(76)
    if (x > 78 and y < 42 and 20 < z < 40) != (x > 78 and y < 42 and 20 < y < 40):
        triggered.add(77)
    if (x > 78 and y < 42 and 20 < z < 40) != (x + z > 98 and y < 42 and 20 < z < 40):
        triggered.add(78)
    if (x > y + 30) != (x > y + 10):
        triggered.add(79)
    if (x > y + 30) != (x > y + 20):
        triggered.add(80)
    if (x > y + 30) != (x > y + 40):
        triggered.add(81)
    if (x > y + 30) != (x > y):
        triggered.add(82)
    if (x > y + 30) != (x + 10 > y + 30):
        triggered.add(83)
    if (x > y + 30) != (x > z + 30):
        triggered.add(84)
    if (x > y + 30) != (z > y + 30):
        triggered.add(85)
    if (x > y + 30) != (x > y + z):
        triggered.add(86)
    if (x > y + 30) != (x > y - z):
        triggered.add(87)
    if (x > y + 30) != (x + z > y + 30):
        triggered.add(88)
    if (abs(x - y) < 10) != (abs(x + y) < 10):
        triggered.add(89)
    if (abs(x - y) < 10) != (abs(x - y) < 15):
        triggered.add(90)
    if (abs(x - y) < 10) != (abs(x - y) < 16):
        triggered.add(91)
    if (abs(x - y) < 10) != (abs(x - y) < 17):
        triggered.add(92)
    if (abs(x - y) < 10) != (abs(x - y) < 20):
        triggered.add(93)
    if (abs(x - y) < 10) != (abs(x - z) < 10):
        triggered.add(94)
    if (abs(x - z) < 15) != (abs(x + z) < 15):
        triggered.add(95)
    if (abs(x - z) < 15) != (abs(x - z) < 25):
        triggered.add(96)
    if (abs(x - z) < 15) != (abs(x - z) < 5):
        triggered.add(97)
    if (abs(x - z) < 15) != (abs(x - z) <= 15):
        triggered.add(98)
    if (abs(x - z) < 15) != (abs(x - y) < 15):
        triggered.add(99)
    if (abs(x - z) < 15) != (abs(y - z) < 15):
        triggered.add(100)
    if (45 < x < 70 and 45 < y < 70 and z < 12) != (45 < x < 70 or 45 < y < 70 and z < 12):
        triggered.add(101)
    if (45 < x < 70 and 45 < y < 70 and z < 12) != (45 < x < 70 and 45 < y < 70 or z < 12):
        triggered.add(102)
    if (45 < x < 70 and 45 < y < 70 and z < 12) != (25 < x < 70 and 45 < y < 70 and z < 12):
        triggered.add(103)
    if (45 < x < 70 and 45 < y < 70 and z < 12) != (45 < x < 50 and 45 < y < 70 and z < 12):
        triggered.add(104)
    if (45 < x < 70 and 45 < y < 70 and z < 12) != (45 < x < 70 and 25 < y < 70 and z < 12):
        triggered.add(105)
    if (45 < x < 70 and 45 < y < 70 and z < 12) != (45 < x < 70 and 45 < y < 50 and z < 12):
        triggered.add(106)
    if (45 < x < 70 and 45 < y < 70 and z < 12) != (45 < x < 70 and 45 < y < 70 and z < 10):
        triggered.add(107)
    if (45 < x < 70 and 45 < y < 70 and z < 12) != (45 < x < 70 and 45 < y < 70 and z < 20):
        triggered.add(108)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 70 or 45 < y < 70 and z > 55):
        triggered.add(109)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 70 and 45 < y < 70 or z > 55):
        triggered.add(110)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (25 < x < 70 and 45 < y < 70 and z > 55):
        triggered.add(111)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 60 and 45 < y < 70 and z > 55):
        triggered.add(112)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 70 and 40 < y < 70 and z > 55):
        triggered.add(113)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 70 and 45 < y < 80 and z > 55):
        triggered.add(114)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 70 and 45 < y < 70 and z > 45):
        triggered.add(115)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 70 and 45 < y < 70 and z > 35):
        triggered.add(116)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 70 and 45 < z < 70 and z > 55):
        triggered.add(117)
    if (45 < x < 70 and 45 < y < 70 and z > 55) != (45 < x < 70 and 45 < y < 70 and y > 55):
        triggered.add(118)
    if (x < 22 and y > 92 and 25 < z < 45) != (x < 22 or y > 92 and 25 < z < 45):
        triggered.add(119)
    if (x < 22 and y > 92 and 25 < z < 45) != (x < 22 and y > 92 or 25 < z < 45):
        triggered.add(120)
    if (x < 22 and y > 92 and 25 < z < 45) != (x < 12 and y > 92 and 25 < z < 45):
        triggered.add(121)
    if (x < 22 and y > 92 and 25 < z < 45) != (x < 22 and y > 52 and 25 < z < 45):
        triggered.add(122)
    if (x < 22 and y > 92 and 25 < z < 45) != (x < 22 and y > 92 and 15 < z < 45):
        triggered.add(123)
    if (x < 22 and y > 92 and 25 < z < 45) != (x < 22 and y > 92 and 25 < z < 35):
        triggered.add(124)

    return triggered


#
execute_Tr = execute_Tr

# === DQN ===
class DQNNetwork(nn.Module):
    def __init__(self, action_size=30):
        super(DQNNetwork, self).__init__()
        hidden_dim = EXPERIMENT_CONFIG['HIDDEN_DIM']

        self.fc1 = nn.Linear(3, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)
        self.output = nn.Linear(hidden_dim, action_size)

    def forward(self, x):
        x = torch.relu(self.fc1(x))
        x = torch.relu(self.fc2(x))
        x = torch.relu(self.fc3(x))
        return self.output(x)


# === Path  ===
class PathReplayBuffer:
    """Path """

    def __init__(self, path_idx, capacity=20000):
        self.path_idx = path_idx
        self.capacity = capacity
        self.buffer = deque(maxlen=capacity)
        self.similarities = deque(maxlen=capacity)  # Similarity

    def push(self, state, action, reward, next_state, done, similarity):
        self.buffer.append((state, action, reward, next_state, done))
        self.similarities.append(similarity)

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return None

        batch = random.sample(self.buffer, batch_size)
        state, action, reward, next_state, done = map(np.stack, zip(*batch))

        return (
            torch.FloatTensor(state).to(device),
            torch.LongTensor(action).to(device),
            torch.FloatTensor(reward).to(device),
            torch.FloatTensor(next_state).to(device),
            torch.BoolTensor(done).to(device)
        )

    def get_top_k(self, k=20):
        """Path Top-K"""
        if len(self.buffer) == 0:
            return []

        # buffersimilarities
        samples_with_sim = list(zip(self.buffer, self.similarities))
        samples_with_sim.sort(key=lambda x: x[1], reverse=True)

        # Top-K
        top_k = samples_with_sim[:k]

        results = []
        for (state, _, _, _, _), similarity in top_k:
            original_state = denormalize_state(state)
            original_state_int = np.round(original_state).astype(int)

            results.append({
                'state': original_state_int,
                'similarity': similarity,
                'triggered': execute_Tr(*original_state_int)
            })

        return results

    def __len__(self):
        return len(self.buffer)


# === DQN(:)===
class ImprovedDQNAgent:
    def __init__(self, num_paths, action_size=30):
        self.action_size = action_size
        self.num_paths = num_paths
        self.epsilon = 0.9
        self.epsilon_min = 0.1
        self.epsilon_decay = 0.995

        self.q_network = DQNNetwork(action_size).to(device)
        self.target_network = DQNNetwork(action_size).to(device)

        lr = EXPERIMENT_CONFIG['LEARNING_RATE']
        self.optimizer = optim.Adam(self.q_network.parameters(), lr=lr)

        # Path
        capacity = EXPERIMENT_CONFIG['REPLAY_BUFFER_CAPACITY']
        self.replay_buffers = {}
        for path_idx in range(num_paths):
            self.replay_buffers[path_idx] = PathReplayBuffer(path_idx, capacity)

        self.replay_train_count = 0
        self.update_target_network()

    def discrete_to_action_delta(self, action_idx):
        #
        delta_values = [5, 3, 2, 1, 0.5, -0.5, -1, -2, -3, -5]

        if action_idx >= 30:
            action_idx = action_idx % 30

        dim = action_idx // 10
        delta_idx = action_idx % 10
        delta = delta_values[delta_idx]

        action_delta = np.zeros(3)
        action_delta[dim] = delta

        return action_delta

    def get_action(self, state):
        mins, maxs = get_bounds()
        normalized_state = (state - (mins + maxs) / 2) / ((maxs - mins) / 2)

        if random.random() < self.epsilon:
            action_idx = random.randint(0, self.action_size - 1)
        else:
            state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)
            with torch.no_grad():
                q_values = self.q_network(state_tensor)
                action_idx = q_values.argmax().item()

        action_delta = self.discrete_to_action_delta(action_idx)
        return action_delta, action_idx

    def store_experience(self, path_idx, state, action_idx, reward, next_state, done, similarity):
        """Path """
        mins, maxs = get_bounds()
        normalized_state = (state - (mins + maxs) / 2) / ((maxs - mins) / 2)
        normalized_next_state = (next_state - (mins + maxs) / 2) / ((maxs - mins) / 2)

        self.replay_buffers[path_idx].push(
            normalized_state, action_idx, reward,
            normalized_next_state, done, similarity
        )

    def replay_train(self, path_idx):
        """Path """
        batch_size = EXPERIMENT_CONFIG['REPLAY_BATCH_SIZE']
        batch = self.replay_buffers[path_idx].sample(batch_size)

        if batch is None:
            return

        states, actions, rewards, next_states, dones = batch

        current_q_values = self.q_network(states).gather(1, actions.unsqueeze(1))

        with torch.no_grad():
            next_q_values = self.target_network(next_states).max(1)[0]
            target_q_values = rewards + (0.99 * next_q_values * ~dones)

        loss = F.mse_loss(current_q_values.squeeze(), target_q_values)

        self.optimizer.zero_grad()
        loss.backward()
        torch.nn.utils.clip_grad_norm_(self.q_network.parameters(), 1.0)
        self.optimizer.step()

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

        self.replay_train_count += 1

        if self.replay_train_count % 2 == 0:
            self.update_target_network()
            print(f"    ->  (Run {self.replay_train_count})")

    def update_target_network(self):
        self.target_network.load_state_dict(self.q_network.state_dict())

    def get_all_top_k(self, k=20):
        """Path Top-K"""
        results = {}
        for path_idx in range(self.num_paths):
            results[path_idx] = self.replay_buffers[path_idx].get_top_k(k)
        return results

    def get_buffer_stats(self):
        """"""
        stats = {}
        for path_idx in range(self.num_paths):
            stats[path_idx] = len(self.replay_buffers[path_idx])
        return stats


# === Metric ===
def calculate_run_performance(run_idx, dqn_results, training_time, total_steps, update_count, agent):
    """ runMetric"""
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    # 1. (Total Reward)
    total_reward = 0
    # 2. (Average Reward)
    average_reward = 0
    # 5. (Convergence)
    convergence = 0
    # 12. (Environment Adaptability)
    environment_adaptability = 0
    # 13. (Generalization Ability)
    generalization_ability = 0
    # 15. (Computational Efficiency)
    computational_efficiency = 0
    # 16. (Policy Update Frequency)
    policy_update_frequency = 0

    # Similarity
    all_similarities = []

    # Metric
    total_samples = 0
    all_rewards = []

    for path_idx in range(num_paths):
        samples = dqn_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']

            total_reward += reward
            all_rewards.append(reward)
            all_similarities.append(similarity)
            total_samples += 1

    # 1.
    total_reward = total_reward

    # 2.
    if total_samples > 0:
        average_reward = total_reward / total_samples

    # 5. (Average Similarity)
    if all_similarities:
        convergence = np.mean(all_similarities)

    # 12. (Similarity)
    if len(all_similarities) > 1:
        environment_adaptability = 1 / (np.std(all_similarities) + 1e-8)

    # 13. (Average Similarity)
    generalization_ability = convergence

    # 15. (/ seconds)
    if training_time > 0:
        computational_efficiency = total_steps / training_time

    # 16.
    if training_time > 0:
        policy_update_frequency = update_count / training_time

    # Similarity
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0

    return {
        'Run': run_idx + 1,

        # Metric
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),

        # Similarity
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }


# === Excel ===
def export_to_excel(all_dqn_results, all_performance_data, target_paths, output_path="DQN_20 run.xlsx"):
    """20 runDQNExcel"""
    print("\nExcel...")

    #
    all_dqn_summary_data = []
    all_dqn_detailed_data = []

    #  run
    for run_idx, (dqn_results, performance_data) in enumerate(zip(all_dqn_results, all_performance_data)):
        # ===== Sheet1: DQNPath  =====
        dqn_summary_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = dqn_results[path_idx]

            if len(samples) == 0:
                dqn_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue

            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = 'Yes' if perfect_count > 0 else 'No'

            dqn_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })
        all_dqn_summary_data.extend(dqn_summary_data)

        # ===== Sheet2: DQNDetailed Sample Data =====
        dqn_detailed_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = dqn_results[path_idx]

            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']

                dqn_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Sample ID': sample_idx + 1,
                    'X': int(state[0]),
                    'Y': int(state[1]),
                    'Z': int(state[2]),
                    'Similarity': round(similarity, 4),
                    'Perfect Coverage': 'Yes' if abs(similarity - 1.0) < 0.001 else 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path))),
                    'Triggered Rules': ', '.join(map(str, sorted(triggered))),
                    'Hit Rule Count': len(target_path.intersection(triggered)),
                    'Target Rule Count': len(target_path)
                })
        all_dqn_detailed_data.extend(dqn_detailed_data)

    # Excel
    dqn_summary_df = pd.DataFrame(all_dqn_summary_data)
    dqn_detailed_df = pd.DataFrame(all_dqn_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet1: DQNPath
        dqn_summary_df.to_excel(writer, sheet_name='DQNPath ', index=False)

        # Sheet2: DQNDetailed Sample Data
        dqn_detailed_df.to_excel(writer, sheet_name='DQNDetailed Sample Data', index=False)

        # Sheet3: Metric
        selected_columns = [
            'Run',
            'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df_selected = performance_df[selected_columns]
        performance_df_selected.to_excel(writer, sheet_name='Metric', index=False)

        #
        workbook = writer.book

        #
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        perfect_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  #

        # === Sheet1 ===
        ws1 = writer.sheets['DQNPath ']
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        #
        for row_idx in range(2, ws1.max_row + 1):
            if ws1.cell(row_idx, 9).value == '':  # Run 9""
                for col_idx in range(1, ws1.max_column + 1):
                    ws1.cell(row_idx, col_idx).fill = perfect_fill

        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 15
        ws1.column_dimensions['F'].width = 15
        ws1.column_dimensions['G'].width = 15
        ws1.column_dimensions['H'].width = 15
        ws1.column_dimensions['I'].width = 15
        ws1.column_dimensions['J'].width = 50

        # === Sheet2 ===
        ws2 = writer.sheets['DQNDetailed Sample Data']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 15
        ws2.column_dimensions['I'].width = 40
        ws2.column_dimensions['J'].width = 40
        ws2.column_dimensions['K'].width = 15
        ws2.column_dimensions['L'].width = 15

        # === Sheet3 ===
        ws3 = writer.sheets['Metric']
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        #
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        for col in columns:
            ws3.column_dimensions[col].width = 18

    print(f"Excel: {output_path}")
    print(f"  - Sheet1: DQNPath  ({len(all_dqn_summary_data)})")
    print(f"  - Sheet2: DQNDetailed Sample Data ({len(all_dqn_detailed_data)})")
    print(f"  - Sheet3: Metric ({len(all_performance_data)})")


# === DQN training(:)===
def train_dqn_workflow():
    print("=" * 80)
    print("DQN training ()")
    print("Similarity:  / target paths")
    print(
        f": Path {EXPERIMENT_CONFIG['NUM_ROUNDS']},{EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(f": Path ,={EXPERIMENT_CONFIG['REPLAY_BUFFER_CAPACITY']}")
    print("=" * 80)

    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    # (Number of Paths)
    agent = ImprovedDQNAgent(num_paths=num_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n: Path {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']}")
    path_samples = {}
    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            state = np.array([
                random.randint(EXPERIMENT_CONFIG['MIN_X'], EXPERIMENT_CONFIG['MAX_X']),
                random.randint(EXPERIMENT_CONFIG['MIN_Y'], EXPERIMENT_CONFIG['MAX_Y']),
                random.randint(EXPERIMENT_CONFIG['MIN_Z'], EXPERIMENT_CONFIG['MAX_Z'])
            ], dtype=np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  Path  {path_idx + 1}/{num_paths}:  {len(samples)} ")

    #
    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size
    num_rounds = EXPERIMENT_CONFIG['NUM_ROUNDS']

    print(f"\n:")
    print(f"  - : {batch_size}")
    print(f"  - Path : {num_batches}")
    print(f"  - Path : {num_rounds}")
    print(f"  - : {EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(
        f"  - : {num_paths} Path  x {num_rounds}  x {num_batches}  = {num_paths * num_rounds * num_batches} ")
    print("-" * 80)

    # :completedPath ,Path
    for path_idx in range(num_paths):
        target_path = target_paths[path_idx]
        print(f"\n{'=' * 80}")
        print(f"Start training path  {path_idx + 1}/{num_paths}")
        print(f": {sorted(target_path)}")
        print(f": replay_buffers[{path_idx}]")
        print(f"{'=' * 80}")

        # Path NUM_ROUNDS
        for round_idx in range(num_rounds):
            print(f"\n{'' * 80}")
            print(f"Path  {path_idx + 1} - Run  {round_idx + 1}/{num_rounds} ")
            print(f"{'' * 80}")

            # Per roundnum_batches
            for batch_idx in range(num_batches):
                print(f"\n   {batch_idx + 1}/{num_batches} (Path {path_idx + 1}, Run {round_idx + 1})")

                #
                batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]

                batch_rewards = []
                batch_similarities = []

                #
                for sample_idx, initial_state in enumerate(batch_samples):
                    state = initial_state.copy()
                    episode_reward = 0
                    final_similarity = 0

                    # STEPS_PER_SAMPLE
                    for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                        action_delta, action_idx = agent.get_action(state)

                        next_state = state + action_delta
                        next_state = clip_state(next_state)

                        triggered = execute_Tr(*next_state)  #
                        reward = unified_reward_function(triggered, target_path)
                        similarity = coverage_similarity(triggered, target_path)

                        done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)

                        # Path
                        agent.store_experience(
                            path_idx, state, action_idx, reward, next_state, done, similarity
                        )

                        state = next_state
                        episode_reward += reward
                        final_similarity = similarity
                        total_steps += 1

                    batch_rewards.append(episode_reward)
                    batch_similarities.append(final_similarity)

                #
                avg_reward = np.mean(batch_rewards)
                avg_similarity = np.mean(batch_similarities)
                max_similarity = np.max(batch_similarities)
                print(f"    ={avg_reward:.2f}, Average Similarity={avg_similarity:.4f}, "
                      f"Similarity={max_similarity:.4f}, epsilon={agent.epsilon:.3f}")

                # Path
                print(f"    (Path {path_idx})...")
                agent.replay_train(path_idx)

                # Path
                buffer_size = len(agent.replay_buffers[path_idx])
                print(f"    Path {path_idx}: {buffer_size}, : {agent.replay_train_count}")

    training_time = time.time() - start_time

    print("\n" + "=" * 80)
    print(f"DQN trainingcompleted! Total elapsed time: {training_time:.2f} seconds, : {total_steps}")
    print(f": {agent.replay_train_count}")
    print(f": {agent.replay_train_count // 2}")

    # Path
    print("\nPath :")
    buffer_stats = agent.get_buffer_stats()
    for path_idx, size in buffer_stats.items():
        print(f"  Path {path_idx + 1}: {size} ")

    print("=" * 80)

    # Top-K
    print(f"\nPath SimilarityMaximum{EXPERIMENT_CONFIG['TOP_K_SAMPLES']}...")
    dqn_top_k_results = agent.get_all_top_k(EXPERIMENT_CONFIG['TOP_K_SAMPLES'])

    return agent, dqn_top_k_results, training_time, total_steps, agent.replay_train_count


# ===  ===
def main():
    print("\n" + "=" * 80)
    print("DQN - 20 run")
    print("Metric")
    print("=" * 80)

    all_dqn_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    # 20
    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'='*80}")
        print(f"Start run  {run_idx + 1}/{EXPERIMENT_CONFIG['NUM_RUNS']}  run")
        print(f"{'='*80}")

        # DQN training
        dqn_agent, dqn_results, training_time, total_steps, update_count = train_dqn_workflow()

        # Metric
        performance_data = calculate_run_performance(
            run_idx, dqn_results, training_time, total_steps, update_count, dqn_agent
        )

        #
        all_dqn_results.append(dqn_results)
        all_performance_data.append(performance_data)

        print(f"\nRun {run_idx + 1} completed!")
        print(f"  Total Reward: {performance_data['Total Reward']}")
        print(f"  Average Reward: {performance_data['Average Reward']}")
        print(f"  Convergence: {performance_data['Convergence']}")

    # Excel(20 run)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"DQN_20 run_{timestamp}.xlsx"
    export_to_excel(all_dqn_results, all_performance_data, target_paths, output_path)

    #
    print("\n" + "=" * 80)
    print("20 run")
    print("=" * 80)

    # Metric
    # Metric Extraction
    total_rewards = [p['Total Reward'] for p in all_performance_data]
    average_rewards = [p['Average Reward'] for p in all_performance_data]
    convergences = [p['Convergence'] for p in all_performance_data]
    environment_adaptabilities = [p['Environment Adaptability'] for p in all_performance_data]
    generalization_abilities = [p['Generalization Ability'] for p in all_performance_data]
    computational_efficiencies = [p['Computational Efficiency'] for p in all_performance_data]
    policy_update_frequencies = [p['Policy Update Frequency'] for p in all_performance_data]
    avg_similarities = [p['Average Similarity'] for p in all_performance_data]

    print(f":")
    print(f"  : {np.mean(total_rewards):.2f}")
    print(f"  Standard deviation: {np.std(total_rewards):.2f}")

    print(f"\n:")
    print(f"  : {np.mean(average_rewards):.4f}")
    print(f"  Standard deviation: {np.std(average_rewards):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(convergences):.4f}")
    print(f"  Standard deviation: {np.std(convergences):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(environment_adaptabilities):.4f}")
    print(f"  Standard deviation: {np.std(environment_adaptabilities):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(generalization_abilities):.4f}")
    print(f"  Standard deviation: {np.std(generalization_abilities):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(computational_efficiencies):.2f}")
    print(f"  Standard deviation: {np.std(computational_efficiencies):.2f}")

    print(f"\n:")
    print(f"  : {np.mean(policy_update_frequencies):.4f}")
    print(f"  Standard deviation: {np.std(policy_update_frequencies):.4f}")

    print(f"\nAverage similarity statistics:")
    print(f"  : {np.mean(avg_similarities):.4f}")
    print(f"  Standard deviation: {np.std(avg_similarities):.4f}")

    print("\n" + "=" * 80)
    print(f" {EXPERIMENT_CONFIG['NUM_RUNS']} completed!")
    print("=" * 80)


if __name__ == "__main__":
    main()