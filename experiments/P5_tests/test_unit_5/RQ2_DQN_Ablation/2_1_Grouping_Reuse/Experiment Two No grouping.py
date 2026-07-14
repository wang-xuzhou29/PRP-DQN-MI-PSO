import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === ()===
MIN_X = 1
MAX_X = 100
MIN_Y = 1
MAX_Y = 400
MIN_Z = 1
MAX_Z = 1000

# === / ===
def normalize_state(state):
    """将状态归一化到 [0, 1] 区间"""
    weather_norm = (state[0] - MIN_X) / (MAX_X - MIN_X)
    time_norm = (state[1] - MIN_Y) / (MAX_Y - MIN_Y)
    z_norm = (state[2] - MIN_Z) / (MAX_Z - MIN_Z)
    return (weather_norm, time_norm, z_norm)

def denormalize_state(state_norm):
    """将归一化状态还原"""
    weather = int(round(state_norm[0] * (MAX_X - MIN_X) + MIN_X))
    time_period = int(round(state_norm[1] * (MAX_Y - MIN_Y) + MIN_Y))
    z = int(round(state_norm[2] * (MAX_Z - MIN_Z) + MIN_Z))

    # 边界保护
    weather = np.clip(weather, MIN_X, MAX_X)
    time_period = np.clip(time_period, MIN_Y, MAX_Y)
    z = np.clip(z, MIN_Z, MAX_Z)

    return (weather, time_period, z)

def normalize_value(value, min_val, max_val):
    """"""
    return (value - min_val) / (max_val - min_val)

def denormalize_value(value_norm, min_val, max_val):
    """"""
    return int(round(value_norm * (max_val - min_val) + min_val))

# ===  ===
def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

# === reward function ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    """"""
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10

    if target_path.issubset(triggered):
        reward += 1

    if prev_triggered is not None:
        prev_sim = jaccard_similarity(prev_triggered, target_path)
        improvement = sim - prev_sim
        reward += improvement * 5

    return reward

def execute_Tr(special_event, emergency_vehicles, visibility):
    triggered = set()

    if (special_event < 50) != (special_event < 70):
        triggered.add(1)
    if (special_event > 50) != (special_event > 30):
        triggered.add(2)
    if (special_event > 50) != (special_event > 70):
        triggered.add(3)
    if (special_event <= 25) != (special_event <= 75):
        triggered.add(4)
    if (special_event >= 25) != (special_event >= 75):
        triggered.add(5)
    if (special_event <= 50) != (special_event <= 25):
        triggered.add(6)

    if (emergency_vehicles < 200) != (emergency_vehicles < 100):
        triggered.add(7)
    if (emergency_vehicles < 200) != (emergency_vehicles < 300):
        triggered.add(8)
    if (emergency_vehicles > 200) != (emergency_vehicles > 100):
        triggered.add(9)
    if (emergency_vehicles > 200) != (emergency_vehicles > 300):
        triggered.add(10)
    if (emergency_vehicles <= 150) != (emergency_vehicles <= 250):
        triggered.add(11)
    if (emergency_vehicles >= 150) != (emergency_vehicles >= 250):
        triggered.add(12)
    if (emergency_vehicles % 2 == 0) != (emergency_vehicles % 3 == 0):
        triggered.add(13)

    if (visibility < 500) != (visibility < 300):
        triggered.add(14)
    if (visibility < 500) != (visibility < 700):
        triggered.add(15)
    if (visibility > 500) != (visibility > 300):
        triggered.add(16)
    if (visibility > 500) != (visibility > 700):
        triggered.add(17)
    if (visibility <= 400) != (visibility <= 600):
        triggered.add(18)
    if (visibility >= 400) != (visibility >= 600):
        triggered.add(19)
    if (visibility % 10 < 5) != (visibility % 20 < 5):
        triggered.add(20)

    if (special_event < 50 and emergency_vehicles < 200) != (
            special_event < 30 and emergency_vehicles < 200):
        triggered.add(21)
    if (special_event < 50 and emergency_vehicles < 200) != (
            special_event < 50 and emergency_vehicles < 150):
        triggered.add(22)
    if (special_event > 50 and emergency_vehicles > 200) != (
            special_event > 30 and emergency_vehicles > 200):
        triggered.add(23)
    if (special_event > 50 and emergency_vehicles > 200) != (
            special_event > 50 and emergency_vehicles > 150):
        triggered.add(24)
    if (special_event <= 40 or emergency_vehicles <= 180) != (
            special_event <= 60 or emergency_vehicles <= 180):
        triggered.add(25)
    if (special_event <= 40 or emergency_vehicles <= 180) != (
            special_event <= 40 or emergency_vehicles <= 320):
        triggered.add(26)

    if (special_event < 50 and visibility < 500) != (special_event < 30 and visibility < 500):
        triggered.add(27)
    if (special_event < 50 and visibility < 500) != (special_event < 50 and visibility < 700):
        triggered.add(28)
    if (special_event > 50 and visibility > 500) != (special_event > 30 and visibility > 500):
        triggered.add(29)
    if (special_event > 50 and visibility > 500) != (special_event > 50 and visibility > 200):
        triggered.add(30)
    if (special_event <= 60 or visibility <= 600) != (special_event <= 40 or visibility <= 600):
        triggered.add(31)
    if (special_event <= 60 or visibility <= 600) != (special_event <= 60 or visibility <= 400):
        triggered.add(32)

    if (emergency_vehicles < 200 and visibility < 500) != (
            emergency_vehicles < 150 and visibility < 500):
        triggered.add(33)
    if (emergency_vehicles < 200 and visibility < 500) != (
            emergency_vehicles < 200 and visibility < 800):
        triggered.add(34)
    if (emergency_vehicles > 200 and visibility > 500) != (
            emergency_vehicles > 150 and visibility > 500):
        triggered.add(35)
    if (emergency_vehicles > 200 and visibility > 500) != (
            emergency_vehicles > 200 and visibility > 200):
        triggered.add(36)
    if (emergency_vehicles <= 350 or visibility <= 700) != (
            emergency_vehicles <= 180 or visibility <= 700):
        triggered.add(37)
    if (emergency_vehicles <= 250 or visibility <= 700) != (
            emergency_vehicles <= 250 or visibility <= 500):
        triggered.add(38)

    if (special_event < 40 or emergency_vehicles > 300) != (
            special_event < 60 or emergency_vehicles > 300):
        triggered.add(39)
    if (special_event < 40 or emergency_vehicles > 300) != (
            special_event < 40 or emergency_vehicles > 250):
        triggered.add(40)
    if (special_event > 60 or visibility < 400) != (special_event > 40 or visibility < 400):
        triggered.add(41)
    if (special_event > 60 or visibility < 400) != (special_event > 60 or visibility < 300):
        triggered.add(42)
    if (emergency_vehicles < 150 or visibility > 700) != (emergency_vehicles < 200 or visibility > 700):
        triggered.add(43)
    if (emergency_vehicles < 150 or visibility > 700) != (emergency_vehicles < 150 or visibility > 400):
        triggered.add(44)

    if (special_event + emergency_vehicles < 350) != (special_event + emergency_vehicles < 200):
        triggered.add(45)
    if (special_event * 2 < emergency_vehicles) != (special_event * 3 < emergency_vehicles):
        triggered.add(46)
    if (visibility - special_event > 400) != (visibility - special_event > 500):
        triggered.add(47)
    if (emergency_vehicles % 50 == 0) != (emergency_vehicles % 100 == 0):
        triggered.add(48)
    if (special_event + visibility > 600) != (special_event + visibility > 500):
        triggered.add(49)
    if (abs(special_event - 50) < 20) != (abs(special_event - 50) < 30):
        triggered.add(50)

    if (special_event < 60 and emergency_vehicles < 250 and visibility > 400) != (
            special_event < 40 and emergency_vehicles < 250 and visibility > 400):
        triggered.add(51)
    if (special_event < 60 and emergency_vehicles < 250 and visibility > 400) != (
            special_event < 60 and emergency_vehicles < 200 and visibility > 400):
        triggered.add(52)
    if (special_event < 60 and emergency_vehicles < 250 and visibility > 400) != (
            special_event < 60 and emergency_vehicles < 250 and visibility > 300):
        triggered.add(53)

    if (special_event > 40 and emergency_vehicles > 150 and visibility < 600) != (
            special_event > 60 and emergency_vehicles > 150 and visibility < 600):
        triggered.add(54)
    if (special_event > 40 and emergency_vehicles > 150 and visibility < 600) != (
            special_event > 40 and emergency_vehicles > 200 and visibility < 600):
        triggered.add(55)
    if (special_event > 40 and emergency_vehicles > 150 and visibility < 600) != (
            special_event > 40 and emergency_vehicles > 150 and visibility < 500):
        triggered.add(56)

    if (special_event < 30 or emergency_vehicles < 100 or visibility > 800) != (
            special_event < 50 or emergency_vehicles < 100 or visibility > 800):
        triggered.add(57)
    if (special_event < 30 or emergency_vehicles < 100 or visibility > 800) != (
            special_event < 30 or emergency_vehicles < 150 or visibility > 800):
        triggered.add(58)
    if (special_event < 30 or emergency_vehicles < 100 or visibility > 800) != (
            special_event < 30 or emergency_vehicles < 100 or visibility > 700):
        triggered.add(59)

    if (special_event > 70 or emergency_vehicles > 300 or visibility < 200) != (
            special_event > 50 or emergency_vehicles > 300 or visibility < 200):
        triggered.add(60)
    if (special_event > 70 or emergency_vehicles > 300 or visibility < 200) != (
            special_event > 70 or emergency_vehicles > 250 or visibility < 200):
        triggered.add(61)
    if (special_event > 70 or emergency_vehicles > 300 or visibility < 200) != (
            special_event > 70 or emergency_vehicles > 300 or visibility < 300):
        triggered.add(62)

    if ((special_event < 50 and emergency_vehicles < 200) or visibility > 700) != (
            (special_event < 30 and emergency_vehicles < 200) or visibility > 700):
        triggered.add(63)
    if ((special_event < 50 and emergency_vehicles < 200) or visibility > 700) != (
            (special_event < 50 and emergency_vehicles < 150) or visibility > 700):
        triggered.add(64)
    if ((special_event < 50 and emergency_vehicles < 200) or visibility > 700) != (
            (special_event < 50 and emergency_vehicles < 200) or visibility > 400):
        triggered.add(65)

    if ((special_event > 50 or emergency_vehicles > 200) and visibility < 500) != (
            (special_event > 30 or emergency_vehicles > 200) and visibility < 500):
        triggered.add(66)
    if ((special_event > 50 or emergency_vehicles > 200) and visibility < 500) != (
            (special_event > 50 or emergency_vehicles > 150) and visibility < 500):
        triggered.add(67)
    if ((special_event > 50 or emergency_vehicles > 200) and visibility < 500) != (
            (special_event > 50 or emergency_vehicles > 200) and visibility < 400):
        triggered.add(68)

    if (special_event + emergency_vehicles > 300 and visibility < 600) != (
            special_event + emergency_vehicles > 250 and visibility < 600):
        triggered.add(69)
    if (special_event + emergency_vehicles > 300 and visibility < 600) != (
            special_event + emergency_vehicles > 300 and visibility < 400):
        triggered.add(70)

    if (special_event * 4 < visibility and emergency_vehicles > 100) != (
            special_event * 5 < visibility and emergency_vehicles > 100):
        triggered.add(71)
    if (special_event * 4 < visibility and emergency_vehicles > 100) != (
            special_event * 4 < visibility and emergency_vehicles > 150):
        triggered.add(72)

    if (visibility - emergency_vehicles > 200 and special_event > 30) != (
            visibility - emergency_vehicles > 300 and special_event > 30):
        triggered.add(73)
    if (visibility - emergency_vehicles > 200 and special_event > 30) != (
            visibility - emergency_vehicles > 200 and special_event > 50):
        triggered.add(74)

    if (special_event % 3 == 0 and emergency_vehicles > 120 and visibility <= 750) != (
            special_event % 4 == 0 and emergency_vehicles > 120 and visibility <= 750):
        triggered.add(75)
    if (special_event % 3 == 0 and emergency_vehicles > 120 and visibility <= 750) != (
            special_event % 3 == 0 and emergency_vehicles > 180 and visibility <= 750):
        triggered.add(76)
    if (special_event % 3 == 0 and emergency_vehicles > 120 and visibility <= 750) != (
            special_event % 3 == 0 and emergency_vehicles > 120 and visibility <= 650):
        triggered.add(77)

    if (abs(special_event - 50) > 10 and emergency_vehicles < 300 and visibility % 100 < 50) != (
            abs(special_event - 50) > 20 and emergency_vehicles < 300 and visibility % 100 < 50):
        triggered.add(78)
    if (abs(special_event - 50) > 10 and emergency_vehicles < 300 and visibility % 100 < 50) != (
            abs(special_event - 50) > 10 and emergency_vehicles < 250 and visibility % 100 < 50):
        triggered.add(79)
    if (abs(special_event - 50) > 10 and emergency_vehicles < 300 and visibility % 100 < 50) != (
            abs(special_event - 50) > 10 and emergency_vehicles < 300 and visibility % 100 < 30):
        triggered.add(80)

    if (20 <= special_event <= 80 and emergency_vehicles < 250) != (
            30 <= special_event <= 80 and emergency_vehicles < 250):
        triggered.add(81)
    if (20 <= special_event <= 80 and emergency_vehicles < 250) != (
            20 <= special_event <= 70 and emergency_vehicles < 250):
        triggered.add(82)
    if (20 <= special_event <= 80 and emergency_vehicles < 250) != (
            20 <= special_event <= 80 and emergency_vehicles < 200):
        triggered.add(83)

    if (100 <= emergency_vehicles <= 300 and visibility > 400) != (
            150 <= emergency_vehicles <= 300 and visibility > 400):
        triggered.add(84)
    if (100 <= emergency_vehicles <= 300 and visibility > 400) != (
            100 <= emergency_vehicles <= 250 and visibility > 400):
        triggered.add(85)
    if (100 <= emergency_vehicles <= 300 and visibility > 400) != (
            100 <= emergency_vehicles <= 300 and visibility > 300):
        triggered.add(86)

    if (300 <= visibility <= 700 and special_event > 25) != (
            400 <= visibility <= 700 and special_event > 25):
        triggered.add(87)
    if (300 <= visibility <= 700 and special_event > 25) != (
            300 <= visibility <= 600 and special_event > 25):
        triggered.add(88)
    if (300 <= visibility <= 700 and special_event > 25) != (
            300 <= visibility <= 700 and special_event > 35):
        triggered.add(89)

    if ((10 <= special_event <= 40) and (150 <= emergency_vehicles <= 350)) != (
            (20 <= special_event <= 40) and (150 <= emergency_vehicles <= 350)):
        triggered.add(90)
    if ((10 <= special_event <= 40) and (150 <= emergency_vehicles <= 350)) != (
            (10 <= special_event <= 30) and (150 <= emergency_vehicles <= 350)):
        triggered.add(91)
    if ((10 <= special_event <= 40) and (150 <= emergency_vehicles <= 350)) != (
            (10 <= special_event <= 40) and (200 <= emergency_vehicles <= 350)):
        triggered.add(92)
    if ((10 <= special_event <= 40) and (150 <= emergency_vehicles <= 350)) != (
            (10 <= special_event <= 40) and (150 <= emergency_vehicles <= 300)):
        triggered.add(93)

    if ((special_event > 60) and (200 < visibility < 800)) != (
            (special_event > 50) and (200 < visibility < 800)):
        triggered.add(94)
    if ((special_event > 60) and (200 < visibility < 800)) != (
            (special_event > 60) and (300 < visibility < 800)):
        triggered.add(95)
    if ((special_event > 60) and (200 < visibility < 800)) != (
            (special_event > 60) and (200 < visibility < 700)):
        triggered.add(96)

    if (special_event < 45 and (emergency_vehicles > 180 or visibility < 450)) != (
            special_event < 55 and (emergency_vehicles > 180 or visibility < 450)):
        triggered.add(97)
    if (special_event < 45 and (emergency_vehicles > 180 or visibility < 450)) != (
            special_event < 45 and (emergency_vehicles > 220 or visibility < 450)):
        triggered.add(98)
    if (special_event < 45 and (emergency_vehicles > 180 or visibility < 450)) != (
            special_event < 45 and (emergency_vehicles > 180 or visibility < 350)):
        triggered.add(99)

    if (emergency_vehicles > 175 and (special_event < 35 or visibility > 650)) != (
            emergency_vehicles > 225 and (special_event < 35 or visibility > 650)):
        triggered.add(100)
    if (emergency_vehicles > 175 and (special_event < 35 or visibility > 650)) != (
            emergency_vehicles > 175 and (special_event < 45 or visibility > 650)):
        triggered.add(101)
    if (emergency_vehicles > 175 and (special_event < 35 or visibility > 650)) != (
            emergency_vehicles > 175 and (special_event < 35 or visibility > 550)):
        triggered.add(102)

    if (special_event + emergency_vehicles // 2 > 150) != (
            special_event + emergency_vehicles // 3 > 150):
        triggered.add(103)
    if (visibility // 10 + special_event > 80) != (visibility // 15 + special_event > 80):
        triggered.add(104)
    if (emergency_vehicles % 10 + special_event % 10 < 15) != (
            emergency_vehicles % 15 + special_event % 10 < 15):
        triggered.add(105)
    if (emergency_vehicles % 10 + special_event % 10 < 15) != (
            emergency_vehicles % 10 + special_event % 15 < 15):
        triggered.add(106)

    if (max(special_event, emergency_vehicles // 4) > 45) != (
            max(special_event, emergency_vehicles // 5) > 45):
        triggered.add(107)
    if (min(visibility // 10, special_event) < 35) != (min(visibility // 15, special_event) < 35):
        triggered.add(108)
    if (min(visibility // 10, special_event) < 35) != (min(visibility // 10, special_event) < 25):
        triggered.add(109)
    if (special_event > visibility // 20) != (special_event > visibility // 25):
        triggered.add(110)

    if (special_event % 2 == emergency_vehicles % 2) != (special_event % 3 == emergency_vehicles % 2):
        triggered.add(111)
    if (special_event % 2 == emergency_vehicles % 2) != (special_event % 2 == emergency_vehicles % 3):
        triggered.add(112)
    if (visibility % 7 == 0) != (visibility % 11 == 0):
        triggered.add(113)
    if ((special_event + emergency_vehicles) % 5 == 0) != (
            (special_event + emergency_vehicles) % 7 == 0):
        triggered.add(114)

    if (special_event & 1 == visibility & 1) != (special_event & 3 == visibility & 1):
        triggered.add(115)
    if (special_event & 1 == visibility & 1) != (special_event & 1 == visibility & 3):
        triggered.add(116)
    if (emergency_vehicles >> 2 > special_event) != (emergency_vehicles >> 3 > special_event):
        triggered.add(117)
    if ((special_event ^ emergency_vehicles) % 50 < 25) != (
            (special_event ^ emergency_vehicles) % 60 < 25):
        triggered.add(118)
    if ((special_event ^ emergency_vehicles) % 50 < 25) != (
            (special_event ^ emergency_vehicles) % 50 < 30):
        triggered.add(119)

    if (len(str(special_event)) + len(str(emergency_vehicles)) > 4) != (
            len(str(special_event)) + len(str(emergency_vehicles)) > 5):
        triggered.add(120)
    if (sum(int(d) for d in str(visibility)) > 10) != (sum(int(d) for d in str(visibility)) > 12):
        triggered.add(121)
    if (special_event ** 2 % 100 < 50) != (special_event ** 2 % 120 < 50):
        triggered.add(122)
    if (special_event ** 2 % 100 < 50) != (special_event ** 2 % 100 < 60):
        triggered.add(123)

    if (abs(special_event - emergency_vehicles // 4) < 20) != (
            abs(special_event - emergency_vehicles // 5) < 20):
        triggered.add(124)
    if (abs(special_event - emergency_vehicles // 4) < 20) != (
            abs(special_event - emergency_vehicles // 4) < 25):
        triggered.add(125)
    if ((visibility + special_event + emergency_vehicles) % 100 > 50) != (
            (visibility + special_event + emergency_vehicles) % 120 > 50):
        triggered.add(126)

    return triggered


# === target path definitions ===
targetPaths = [
    [2, 4, 5, 6, 8, 10, 11, 12, 15, 17, 18, 19, 20, 23, 28, 29, 44, 45, 49, 52, 57, 65, 74, 83, 89, 91, 98, 101, 102, 104, 105, 107, 111, 112, 113, 115, 117, 118, 120],
    [2, 4, 5, 6, 7, 9, 11, 12, 13, 14, 16, 18, 19, 21, 22, 27, 33, 43, 44, 45, 47, 57, 63, 64, 66, 67, 73, 74, 76, 91, 92, 104, 108, 112, 116, 117, 118, 120, 121],
    [2, 4, 5, 6, 7, 9, 11, 12, 13, 14, 16, 20, 21, 22, 27, 33, 42, 43, 45, 53, 57, 63, 64, 66, 67, 75, 86, 87, 89, 91, 92, 100, 109, 112, 114, 117, 118, 120, 122],
    [2, 4, 5, 6, 7, 9, 11, 12, 13, 14, 16, 18, 19, 21, 22, 27, 33, 43, 44, 47, 49, 57, 63, 64, 66, 67, 74, 76, 91, 92, 104, 108, 112, 114, 117, 118, 120],
    [7, 9, 11, 12, 13, 15, 17, 18, 19, 20, 22, 28, 34, 35, 43, 44, 45, 47, 49, 50, 64, 75, 81, 92, 98, 100, 110, 111, 112, 118, 120, 122, 124, 125, 126],
    [2, 4, 5, 6, 8, 10, 11, 12, 20, 23, 25, 26, 29, 31, 37, 39, 41, 45, 51, 52, 74, 83, 100, 103, 105, 106, 107, 110, 113, 114, 115, 117, 118, 120],
    [2, 4, 5, 6, 8, 9, 11, 12, 13, 15, 16, 18, 19, 28, 44, 45, 47, 49, 52, 57, 65, 73, 74, 78, 83, 91, 98, 101, 104, 107, 108, 112, 117, 120],
    [1, 3, 4, 5, 8, 10, 13, 14, 16, 18, 19, 26, 30, 32, 36, 40, 44, 45, 49, 60, 61, 65, 68, 70, 78, 79, 80, 85, 108, 111, 112, 116, 117, 120],
    [2, 4, 5, 6, 8, 10, 11, 12, 14, 16, 23, 25, 26, 27, 36, 39, 42, 45, 53, 54, 57, 69, 75, 83, 86, 87, 97, 103, 109, 111, 117, 118, 120, 122],
    [1, 3, 4, 5, 8, 10, 11, 12, 14, 16, 25, 26, 30, 36, 39, 42, 45, 53, 54, 60, 69, 75, 83, 86, 87, 94, 103, 104, 109, 111, 113, 117, 120, 121],
    [2, 4, 5, 6, 8, 10, 11, 12, 13, 20, 23, 25, 26, 29, 31, 37, 39, 41, 45, 51, 52, 57, 59, 74, 83, 97, 103, 111, 112, 113, 117, 118, 120, 122],
    [2, 4, 5, 6, 7, 9, 12, 15, 17, 18, 19, 20, 21, 22, 28, 29, 34, 43, 44, 48, 57, 63, 64, 72, 74, 76, 91, 92, 104, 116, 117, 119, 120],
    [1, 3, 4, 5, 8, 10, 11, 12, 14, 16, 25, 26, 30, 36, 39, 42, 45, 53, 54, 60, 69, 75, 83, 86, 87, 94, 97, 103, 104, 105, 106, 108, 120],
    [7, 9, 11, 12, 13, 15, 17, 22, 28, 34, 35, 43, 44, 45, 50, 64, 75, 77, 81, 92, 98, 100, 104, 106, 107, 111, 112, 116, 118, 120, 122, 124, 126],
    [8, 10, 11, 12, 14, 16, 18, 19, 26, 30, 32, 36, 44, 45, 49, 50, 65, 68, 70, 73, 75, 80, 82, 83, 108, 111, 113, 118, 120, 122, 124, 126],
    [4, 5, 6, 7, 9, 11, 12, 13, 14, 16, 18, 19, 22, 33, 43, 44, 50, 64, 67, 81, 89, 92, 99, 109, 112, 113, 115, 117, 119, 120],
    [1, 3, 4, 5, 8, 10, 11, 12, 13, 25, 26, 30, 36, 39, 45, 54, 60, 62, 69, 75, 83, 94, 103, 104, 105, 106, 109, 112, 118, 120],
    [1, 3, 4, 5, 13, 15, 17, 18, 19, 25, 38, 41, 44, 47, 49, 54, 56, 65, 70, 75, 94, 97, 108, 112, 113, 115, 116, 117, 120, 121],
    [7, 9, 11, 12, 14, 16, 18, 19, 24, 30, 32, 33, 43, 44, 45, 49, 55, 65, 68, 69, 71, 73, 75, 80, 103, 108, 111, 113, 118],
    [7, 9, 11, 12, 13, 15, 17, 22, 28, 34, 35, 43, 44, 64, 75, 90, 92, 104, 105, 111, 112, 119, 120, 122, 123, 124, 125, 126],
    [7, 9, 11, 12, 13, 20, 24, 26, 30, 33, 43, 45, 46, 55, 69, 75, 80, 95, 103, 106, 111, 112, 114, 120, 122],
    [7, 9, 11, 12, 24, 34, 35, 45, 46, 50, 59, 75, 76, 77, 82, 96, 103, 105, 106, 113, 115, 120, 121],
    [2, 4, 5, 6, 7, 9, 21, 29, 31, 39, 41, 51, 72, 74, 84, 106, 110, 111, 113, 114, 118, 120, 122],
    [7, 9, 13, 15, 17, 34, 45, 58, 65, 72, 80, 84, 88, 111, 112, 120, 126],
    [4, 5, 6, 13, 37, 48, 50, 93, 103, 109, 112, 114, 120, 122]
]

# 
target_paths = [set(path) for path in targetPaths]

targetPaths = [set(path) for path in target_paths]
NUM_PATHS = len(targetPaths)

def jaccard_similarity(set1, set2):
    """Compute Jaccard similarity"""
    if not set2:
        return 0.0

    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0

# === Path Similarity ===
def compute_path_similarity_matrix(paths):
    """Path Similarity"""
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix

def group_paths_by_similarity(paths, threshold_percentile=50):
    """SimilarityPath """
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.percentile(avg_sim_scores, threshold_percentile)

    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]

    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)

    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group

# === Sample generation ===
def compute_robustness(state, path, sample_size=9):
    """()"""
    base = execute_Tr(state[0], state[1], state[2])
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0

    # : 
    # weather1-6, +/-1
    # time_period1-6, +/-1
    # z1-60, +/-6(10%), +/-3(5%)
    deltas = [
        (-1, -1, -6), (0, -1, 0), (1, -1, 6),
        (-1, 0, -6), (1, 0, 6),
        (-1, 1, -6), (0, 1, 0), (1, 1, 6),
        (0, 0, 0)
    ]

    for dw, dt, dz in deltas[:sample_size]:
        if dw == dt == dz == 0:
            continue

        neighbor_weather = int(np.clip(state[0] + dw, MIN_X, MAX_X))
        neighbor_time = int(np.clip(state[1] + dt, MIN_Y, MAX_Y))
        neighbor_z = int(np.clip(state[2] + dz, MIN_Z, MAX_Z))
        neighbor = (neighbor_weather, neighbor_time, neighbor_z)

        n_trig = execute_Tr(neighbor[0], neighbor[1], neighbor[2])
        if not n_trig:
            continue

        rob += jaccard_similarity(n_trig, base)
        neighbors += 1

    return rob / neighbors if neighbors > 0 else 0.0

def generate_samples_for_all_paths(num_candidates=2000, top_k=200, run_id=1):
    """Path """
    BEST_WEIGHTS = [0.55, 0.25, 0.2]

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_individual.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Individual Path {path_id} (Weighted Screening) - Run {run_id}\n")
            f.write("weather time_period z\tScore\tSimilarity\tRobustness\tLengthDiff\n")
            for s in samples:
                weather, time_period, z = s['state']
                f.write(
                    f"{weather} {time_period} {z}\t{s['score']:.4f}\t{s['similarity']:.4f}\t"
                    f"{s['robustness']:.4f}\t{s['length_diff']:.4f}\n"
                )

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_individual"

    for path_idx in range(len(targetPaths)):
        path = targetPaths[path_idx]
        candidate_samples = []
        attempts = 0
        max_attempts = num_candidates * 10

        while len(candidate_samples) < num_candidates and attempts < max_attempts:
            attempts += 1

            weather = np.random.randint(MIN_X, MAX_X + 1)
            time_period = np.random.randint(MIN_Y, MAX_Y + 1)
            z = np.random.randint(MIN_Z, MAX_Z + 1)
            state = (weather, time_period, z)

            triggered = execute_Tr(weather, time_period, z)
            if not triggered:
                continue

            sim = jaccard_similarity(triggered, path)
            rob = compute_robustness(state, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))

            candidate_samples.append({
                'state': state,
                'similarity': sim,
                'robustness': rob,
                'length_diff': len_diff,
                'triggered': triggered
            })

        if candidate_samples:
            for sample in candidate_samples:
                score = (BEST_WEIGHTS[0] * sample['similarity'] +
                         BEST_WEIGHTS[1] * sample['robustness'] +
                         BEST_WEIGHTS[2] * sample['length_diff'])
                sample['score'] = score

            candidate_samples.sort(key=lambda x: x['score'], reverse=True)
            selected_samples = candidate_samples[:top_k]
            save_samples(path_id=path_idx + 1, samples=selected_samples, base_dir=base_dir)

# === (: )===
class SharedExperienceReplay:
    """()"""

    def __init__(self, capacity=20000):
        self.capacity = capacity
        self.buffer = deque(maxlen=self.capacity)
        self.priorities = deque(maxlen=self.capacity)

    def append(self, experience):
        """"""
        self.buffer.append(experience)
        self.priorities.append(experience[-1])

    def sample(self, batch_size, alpha=0.6):
        """"""
        if len(self.buffer) < batch_size:
            return [], [], []

        priorities = np.array(self.priorities, dtype=np.float64)
        priorities = np.power(priorities, alpha)
        probabilities = priorities / np.sum(priorities)

        batch_indices = np.random.choice(len(self.buffer), batch_size, p=probabilities, replace=False)
        batch = [self.buffer[idx] for idx in batch_indices]

        return batch, batch_indices, probabilities[batch_indices]

    def update_priorities(self, batch_indices, td_errors):
        """"""
        for idx, td_error in zip(batch_indices, td_errors):
            if idx < len(self.priorities):
                self.priorities[idx] = max(abs(td_error), 1e-6)

    def __len__(self):
        return len(self.buffer)

    def get_high_reward_samples(self, target_path, num_samples=20):
        """(, )"""
        if len(self.buffer) == 0:
            return []

        samples_with_scores = []
        seen_states = set()  # 

        for experience in self.buffer:
            state_tensor = experience[0]
            state_norm = state_tensor.cpu().numpy().flatten()

            # 
            state_tuple = denormalize_state((state_norm[0], state_norm[1], state_norm[2]))

            # 
            if state_tuple in seen_states:
                continue
            seen_states.add(state_tuple)

            triggered = execute_Tr(state_tuple[0], state_tuple[1], state_tuple[2])
            reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)

            samples_with_scores.append((state_tuple, reward, sim, triggered))

        samples_with_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_scores[:num_samples]

def load_path_data(file_path):
    """screeningPath """
    path_data = []

    if not os.path.exists(file_path):
        print(f":  {file_path}")
        return path_data

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
            for line in lines[2:]:
                parts = line.strip().split("\t")
                if parts:
                    values = parts[0].split()
                    if len(values) >= 3:
                        state = (int(values[0]), int(values[1]), int(values[2]))
                        path_data.append(state)
    except Exception as e:
        print(f" {file_path}: {e}")

    return path_data

# === DQN ===
class DQN(nn.Module):
    """Q"""

    def __init__(self, state_dim, action_dim, hidden_dims=[128, 64]):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, hidden_dims[0])
        self.fc2 = nn.Linear(hidden_dims[0], hidden_dims[1])
        self.fc3 = nn.Linear(hidden_dims[1], action_dim)
        self.dropout = nn.Dropout(0.1)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = self.dropout(x)
        x = torch.relu(self.fc2(x))
        return self.fc3(x)

# === DQN Agent with PER(: )===
class DQNAgentWithPER:
    """DQN()"""

    def __init__(self, state_dim, action_dim, replay_buffer,
                 gamma=0.99, epsilon=1.0, epsilon_decay=0.995,
                 epsilon_min=0.1, learning_rate=0.001, alpha=0.6, beta=0.4):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer
        self.alpha = alpha
        self.beta = beta

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        """
        ()
        30, : 
        - weather: +/-1, 0(x2)
        - time_period: +/-1, 0(x2)
        - z: +/-12(20%), +/-6(10%), +/-3(5%), 0(x2)
        """
        delta_values_weather_time = [1, 0, 0, -1]  # 
        delta_values_z = [12, 6, 3, 0, 0, -3, -6, -12]  # z

        dim = action_idx // 10
        delta_idx = action_idx % 10

        if dim == 0:  # weather
            if delta_idx >= 4:
                delta_idx = 3
            return (delta_values_weather_time[delta_idx], 0, 0)
        elif dim == 1:  # time_period
            if delta_idx >= 4:
                delta_idx = 3
            return (0, delta_values_weather_time[delta_idx], 0)
        elif dim == 2:  # z
            if delta_idx >= 8:
                delta_idx = 7
            return (0, 0, delta_values_z[delta_idx])

    def act(self, state_norm, legal_actions=None):
        """()"""
        if legal_actions is None:
            legal_actions = list(range(self.action_dim))

        if not legal_actions:
            return None

        if random.random() < self.epsilon:
            return random.choice(legal_actions)

        state_tensor = torch.tensor(state_norm, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)[0]

        legal_q_values = q_values[legal_actions]
        best_legal_idx = torch.argmax(legal_q_values).item()
        return legal_actions[best_legal_idx]

    def get_legal_actions(self, state):
        """()"""
        legal_actions = []

        for action_idx in range(self.action_dim):
            dw, dt, dz = self.decode_action(action_idx)

            next_weather = state[0] + dw
            next_time = state[1] + dt
            next_z = state[2] + dz

            if (MIN_X <= next_weather <= MAX_X and
                    MIN_Y <= next_time <= MAX_Y and
                    MIN_Z <= next_z <= MAX_Z):
                legal_actions.append(action_idx)

        return legal_actions

    def store_transition(self, state, action, reward, next_state, done):
        """()"""
        # 
        state_norm = normalize_state(state)
        next_state_norm = normalize_state(next_state)

        state_tensor = torch.tensor(state_norm, dtype=torch.float32).unsqueeze(0).to(device)
        next_state_tensor = torch.tensor(next_state_norm, dtype=torch.float32).unsqueeze(0).to(device)

        with torch.no_grad():
            q_values = self.model(state_tensor)
            next_q_values = self.target_model(next_state_tensor)
            max_next_q = next_q_values.max(1)[0]
            target_q = reward + (self.gamma * max_next_q * (1 - done))
            td_error = abs(q_values[0][action].item() - target_q.item())

        self.replay_buffer.append((state_tensor, action, reward, next_state_tensor, done, td_error))
        return td_error

    def train(self, batch_size=32):
        """"""
        if len(self.replay_buffer) < batch_size:
            return 0.0

        batch, batch_indices, probabilities = self.replay_buffer.sample(batch_size, alpha=self.alpha)

        if not batch:
            return 0.0

        states, actions, rewards, next_states, dones, _ = zip(*batch)

        weights = (len(self.replay_buffer) * probabilities) ** (-self.beta)
        weights = weights / weights.max()
        weights = torch.tensor(weights, dtype=torch.float32).to(device)

        states = torch.cat(states).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.cat(next_states).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q = self.target_model(next_states).max(1)[0].detach()
        target_q = rewards + (self.gamma * next_max_q * (1 - dones))

        td_errors = current_q - target_q
        weighted_loss = (td_errors.pow(2) * weights).mean()

        self.optimizer.zero_grad()
        weighted_loss.backward()
        torch.nn.utils.clip_grad_norm_(self.model.parameters(), 1.0)
        self.optimizer.step()

        new_priorities = abs(td_errors.detach().cpu().numpy())
        self.replay_buffer.update_priorities(batch_indices, new_priorities)

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

        return weighted_loss.item()

    def update_target_model(self):
        """"""
        self.target_model.load_state_dict(self.model.state_dict())

# ===  ===
def generate_and_train_for_individual_paths(path_documents, repeats=5, batch_size=32, run_id=1):
    """"""
    state_dim = 3
    action_dim = 30

    shared_replay_buffer = SharedExperienceReplay(capacity=20000)
    agent = DQNAgentWithPER(state_dim, action_dim, shared_replay_buffer)

    total_cumulative_reward = 0
    path_rewards = {}

    print(f"\n===  {run_id}/20 Start training(+)===")
    start_time = time.time()

    SAMPLES_PER_BATCH = 50
    NUM_BATCHES = 4
    STEPS_PER_SAMPLE = 3

    for path_idx in range(len(targetPaths)):
        path_id = path_idx + 1
        print(f"\n{'=' * 60}")
        print(f"Path  {path_id}/{NUM_PATHS}")
        print(f"{'=' * 60}")

        file_path = os.path.join(path_documents, f"path{path_id}_individual.txt")
        if not os.path.exists(file_path):
            print(f"  :  {file_path}")
            continue

        path_data = load_path_data(file_path)
        if not path_data:
            print(f"  : Path  {path_id} ")
            continue

        target_path = targetPaths[path_idx]

        if path_idx not in path_rewards:
            path_rewards[path_idx] = 0

        for repeat_idx in range(repeats):
            print(f"\n  Run  {repeat_idx + 1}/{repeats} ")

            for batch_idx in range(NUM_BATCHES):
                batch_start = batch_idx * SAMPLES_PER_BATCH
                batch_end = min(batch_start + SAMPLES_PER_BATCH, len(path_data))

                print(f"     {batch_idx + 1}/{NUM_BATCHES} ( {batch_start}-{batch_end})")

                for sample_idx in range(batch_start, batch_end):
                    state = path_data[sample_idx]
                    prev_state = None
                    prev_triggered = None

                    for step in range(STEPS_PER_SAMPLE):
                        legal_actions = agent.get_legal_actions(state)

                        if not legal_actions:
                            break

                        # 
                        state_norm = normalize_state(state)
                        action = agent.act(state_norm, legal_actions)
                        if action is None:
                            break

                        dw, dt, dz = agent.decode_action(action)

                        next_state = (
                            int(np.clip(state[0] + dw, MIN_X, MAX_X)),
                            int(np.clip(state[1] + dt, MIN_Y, MAX_Y)),
                            int(np.clip(state[2] + dz, MIN_Z, MAX_Z))
                        )

                        triggered = execute_Tr(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered,
                                                prev_triggered, prev_state)
                        done = (step == STEPS_PER_SAMPLE - 1)

                        # 
                        agent.store_transition(state, action, reward, next_state, done)

                        prev_state = state
                        prev_triggered = triggered
                        state = next_state

                        total_cumulative_reward += reward
                        path_rewards[path_idx] += reward

                if len(agent.replay_buffer) >= batch_size:
                    loss = agent.train(batch_size)
                    print(f"       {batch_idx + 1} completed, : {loss:.4f}")

                if (batch_idx + 1) % 2 == 0:
                    agent.update_target_model()
                    print(f"      completed {batch_idx + 1} , ")

        print(f"\nPath  {path_id} completed, : {path_rewards[path_idx]:.2f}")
        print(f"Shared Buffer Size: {len(shared_replay_buffer)}")

    training_time = time.time() - start_time
    print(f"\n===  {run_id}/20 completed, : {training_time:.2f} seconds ===")

    return agent, shared_replay_buffer, total_cumulative_reward, path_rewards, training_time

# === Excel ===
def create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir):
    """Excel"""
    os.makedirs(output_dir, exist_ok=True)

    similar_group_paths = [idx + 1 for idx in similar_group]
    isolated_group_paths = [idx + 1 for idx in isolated_group]

    wb = Workbook()

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    header_color = "4472C4"
    similar_group_color = "E2EFDA"
    isolated_group_color = "FCE4D6"
    stats_color = "FFF2CC"

    # === 1: Path  ===
    ws_paths = wb.active
    ws_paths.title = "Path "

    path_headers = ['Path ID', ''] + [f'Run {i}' for i in range(1, 21)] + \
                   ['Average Similarity', 'Maximum Similarity', 'Minimum Similarity', 'Standard deviation']

    for col, header in enumerate(path_headers, 1):
        cell = ws_paths.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_paths.row_dimensions[1].height = 30

    for path_id in range(1, NUM_PATHS + 1):
        row = path_id + 1

        if path_id in similar_group_paths:
            group_type = "High-correlation path group"
            row_color = similar_group_color
        elif path_id in isolated_group_paths:
            group_type = "Low-correlation path group"
            row_color = isolated_group_color
        else:
            group_type = "Ungrouped"
            row_color = "FFFFFF"

        cell = ws_paths.cell(row=row, column=1, value=f"Path {path_id}")
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        cell.border = thin_border

        cell = ws_paths.cell(row=row, column=2, value=group_type)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        cell.border = thin_border

        path_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            sim = run_data['path_similarities'].get(path_id, {}).get('avg_similarity', 0.0)
            path_similarities.append(sim)

            cell = ws_paths.cell(row=row, column=3 + run_idx, value=round(sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        stats_values = [
            round(np.mean(path_similarities), 4),
            round(np.max(path_similarities), 4),
            round(np.min(path_similarities), 4),
            round(np.std(path_similarities), 4)
        ]

        for i, value in enumerate(stats_values):
            cell = ws_paths.cell(row=row, column=23 + i, value=value)
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
            cell.font = Font(bold=True, size=10)
            cell.border = thin_border

    ws_paths.column_dimensions['A'].width = 13
    ws_paths.column_dimensions['B'].width = 16
    for col in range(3, 23):
        ws_paths.column_dimensions[get_column_letter(col)].width = 10
    for col in range(23, 27):
        ws_paths.column_dimensions[get_column_letter(col)].width = 13

    # === 2:  ===
    ws_groups = wb.create_sheet("")

    group_headers = ['Group Name', 'Included Paths'] + [f'Run {i}' for i in range(1, 21)] + \
                    ['Average Similarity', 'Standard deviation']

    for col, header in enumerate(group_headers, 1):
        cell = ws_groups.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_groups.row_dimensions[1].height = 30

    row = 2

    # High-correlation path group
    cell = ws_groups.cell(row=row, column=1, value="High-correlation path group")
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    cell.border = thin_border

    cell = ws_groups.cell(row=row, column=2, value=','.join(map(str, similar_group_paths)))
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    cell.border = thin_border

    group_similarities = []
    for run_idx, run_data in enumerate(all_runs_data):
        group_sim = np.mean([
            run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
            for p in similar_group_paths
        ])
        group_similarities.append(group_sim)

        cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(group_sim, 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    cell = ws_groups.cell(row=row, column=23, value=round(np.mean(group_similarities), 4))
    cell.number_format = '0.0000'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border

    cell = ws_groups.cell(row=row, column=24, value=round(np.std(group_similarities), 4))
    cell.number_format = '0.0000'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border

    row += 1

    # Low-correlation path group
    if isolated_group_paths:
        cell = ws_groups.cell(row=row, column=1, value="Low-correlation path group")
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        cell.border = thin_border

        cell = ws_groups.cell(row=row, column=2, value=','.join(map(str, isolated_group_paths)))
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        cell.border = thin_border

        isolated_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            iso_sim = np.mean([
                run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
                for p in isolated_group_paths
            ])
            isolated_similarities.append(iso_sim)

            cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(iso_sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        cell = ws_groups.cell(row=row, column=23, value=round(np.mean(isolated_similarities), 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border

        cell = ws_groups.cell(row=row, column=24, value=round(np.std(isolated_similarities), 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border

    ws_groups.column_dimensions['A'].width = 16
    ws_groups.column_dimensions['B'].width = 22
    for col in range(3, 23):
        ws_groups.column_dimensions[get_column_letter(col)].width = 10
    ws_groups.column_dimensions[get_column_letter(23)].width = 14
    ws_groups.column_dimensions[get_column_letter(24)].width = 12

    # === 3: Detailed Sample Data ===
    ws_samples = wb.create_sheet("Detailed Sample Data")

    sample_headers = ['Run', 'Path ID', 'Sample ID', '', '',
                      '', 'Similarity', 'Triggered Rule Set']

    for col, header in enumerate(sample_headers, 1):
        cell = ws_samples.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_samples.row_dimensions[1].height = 30

    sample_row = 2
    for run_idx, run_data in enumerate(all_runs_data, 1):
        for path_id in range(1, NUM_PATHS + 1):
            samples = run_data['path_samples'].get(path_id, [])

            if path_id in similar_group_paths:
                path_color = similar_group_color
            elif path_id in isolated_group_paths:
                path_color = isolated_group_color
            else:
                path_color = "FFFFFF"

            for sample_idx, (state_tuple, reward, sim, triggered) in enumerate(samples, 1):
                weather, time_period, z = state_tuple
                triggered_str = ','.join(map(str, sorted(triggered)))

                cell = ws_samples.cell(row=sample_row, column=1, value=f"Run {run_idx}")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=path_color, end_color=path_color, fill_type="solid")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=2, value=f"Path {path_id}")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=path_color, end_color=path_color, fill_type="solid")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=3, value=sample_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                for col_offset, value in enumerate([weather, time_period, z]):
                    cell = ws_samples.cell(row=sample_row, column=4 + col_offset, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=7, value=round(sim, 4))
                cell.number_format = '0.0000'
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=8, value=f"{{{triggered_str}}}")
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border

                sample_row += 1

    sample_widths = [13, 13, 11, 10, 10, 12, 12, 45]
    for i, width in enumerate(sample_widths, 1):
        ws_samples.column_dimensions[get_column_letter(i)].width = width

    # === 4: Run Statistics Summary ===
    ws_summary = wb.create_sheet("Run Statistics Summary")

    summary_headers = ['Run', 'Training Time( seconds)', 'Overall Average Similarity', 'Maximum Similarity',
                       'Minimum Similarity', '', '', 'Shared Buffer Size']

    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_summary.row_dimensions[1].height = 30

    for run_idx, run_data in enumerate(all_runs_data, 1):
        row = run_idx + 1

        # Average Similarity
        high_group_avg = np.mean([
            run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
            for p in similar_group_paths
        ])

        low_group_avg = 0.0
        if isolated_group_paths:
            low_group_avg = np.mean([
                run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
                for p in isolated_group_paths
            ])

        values = [
            f"Run {run_idx}",
            round(run_data['training_time'], 2),
            round(run_data['overall_avg_similarity'], 4),
            round(run_data['max_similarity'], 4),
            round(run_data['min_similarity'], 4),
            round(high_group_avg, 4),
            round(low_group_avg, 4),
            20000  # 
        ]

        for col, value in enumerate(values, 1):
            cell = ws_summary.cell(row=row, column=col, value=value)
            if col == 1:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 2:
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col >= 3 and col <= 7:
                cell.number_format = '0.0000'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # 
    stat_row = len(all_runs_data) + 2
    stat_labels = ['', '/', '', '', '', '', '', '']

    for col, label in enumerate(stat_labels, 1):
        cell = ws_summary.cell(row=stat_row, column=col, value=label)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    stat_row += 1

    # 
    training_times = [r['training_time'] for r in all_runs_data]
    overall_avgs = [r['overall_avg_similarity'] for r in all_runs_data]
    max_sims = [r['max_similarity'] for r in all_runs_data]
    min_sims = [r['min_similarity'] for r in all_runs_data]

    high_group_avgs = []
    low_group_avgs = []
    for run_data in all_runs_data:
        high_avg = np.mean([
            run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
            for p in similar_group_paths
        ])
        high_group_avgs.append(high_avg)

        if isolated_group_paths:
            low_avg = np.mean([
                run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
                for p in isolated_group_paths
            ])
            low_group_avgs.append(low_avg)

    stat_values = [
        '',
        round(np.sum(training_times), 2),
        round(np.mean(overall_avgs), 4),
        round(np.max(max_sims), 4),
        round(np.min(min_sims), 4),
        round(np.mean(high_group_avgs), 4),
        round(np.mean(low_group_avgs), 4) if low_group_avgs else 0.0,
        20000
    ]

    for col, value in enumerate(stat_values, 1):
        cell = ws_summary.cell(row=stat_row, column=col, value=value)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        if col == 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col == 2:
            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col >= 3 and col <= 7:
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    summary_widths = [13, 16, 18, 14, 14, 16, 16, 14]
    for i, width in enumerate(summary_widths, 1):
        ws_summary.column_dimensions[get_column_letter(i)].width = width

    output_path = os.path.join(output_dir, "20 run_.xlsx")
    wb.save(output_path)
    print(f"\n Consolidated Excel report generated: {output_path}")
    print(f"   4: Path , , Detailed Sample Data, Run Statistics Summary")

def run_20_times_training():
    """20"""
    model_path_base = r"D:\Experiment\CNN\DQNNEW\saved_models_new_vars"
    path_documents = r"D:\Experiment\CNN\DQNNEW\path_samples_individual"
    output_dir = r"D:\Experiment\CNN\ComparisonExperiment2\excel_reports_new_vars"

    os.makedirs(model_path_base, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    similar_group, isolated_group = group_paths_by_similarity(targetPaths)

    similar_group_display = [idx + 1 for idx in similar_group]
    isolated_group_display = [idx + 1 for idx in isolated_group]

    print("=" * 60)
    print("20 - ")
    print("=" * 60)
    print(f"\nPath : {NUM_PATHS} ")
    print(f"\nAutomatic grouping results:")
    print(f"  Similar path group: {similar_group_display}")
    print(f"  Isolated path group: {isolated_group_display}")
    print(f"\n:")
    print(f"   (weather): {MIN_X}-{MAX_X}")
    print(f"   (time_period): {MIN_Y}-{MAX_Y}")
    print(f"   (z): {MIN_Z}-{MAX_Z}")
    print(f"\n:")
    print(f"  x = z ()")
    print(f"  y = (weather * time_period * 10 + z) % 100 + 1")
    print(f"\n:")
    print(f"  : [0, 1]")
    print(f"  : ")
    print(f"  : (seen_states)")
    print(f"\n (30):")
    print(f"  weather: +/-1, 0(x2)")
    print(f"  time_period: +/-1, 0(x2)")
    print(f"  z: +/-12, +/-6, +/-3, 0(x2)")
    print("\n" + "=" * 60)

    all_runs_data = []
    total_start_time = time.time()

    for run_id in range(1, 21):
        print(f"\n{'=' * 60}")
        print(f"Start run  {run_id}/20  run")
        print(f"{'=' * 60}")

        print(f"[Run {run_id}] Generating samples...")
        generate_samples_for_all_paths(num_candidates=2000, top_k=200, run_id=run_id)

        print(f"[Run {run_id}] Start training...")
        agent, shared_buffer, total_cumulative_reward, path_rewards, training_time = \
            generate_and_train_for_individual_paths(path_documents, repeats=5,
                                                    batch_size=32, run_id=run_id)

        model_path = os.path.join(model_path_base, f"trained_model_run_{run_id}.pth")
        torch.save({
            'model_state_dict': agent.model.state_dict(),
            'optimizer_state_dict': agent.optimizer.state_dict(),
            'epsilon': agent.epsilon,
            'run_id': run_id,
            'normalized': True,
            'value_ranges': {
                'weather': [MIN_X, MAX_X],
                'time_period': [MIN_Y, MAX_Y],
                'z': [MIN_Z, MAX_Z]
            }
        }, model_path)
        print(f"[Run {run_id}] Model saved: {model_path}")

        run_data = {
            'run_id': run_id,
            'training_time': training_time,
            'total_reward': total_cumulative_reward,
            'path_rewards': path_rewards,
            'path_similarities': {},
            'path_samples': {}
        }

        all_similarities = []
        for path_idx in range(len(targetPaths)):
            target_path = targetPaths[path_idx]
            high_reward_samples = shared_buffer.get_high_reward_samples(target_path, num_samples=20)

            if high_reward_samples:
                similarities = [sim for _, _, sim, _ in high_reward_samples]
                run_data['path_similarities'][path_idx + 1] = {
                    'avg_similarity': np.mean(similarities),
                    'max_similarity': np.max(similarities),
                    'min_similarity': np.min(similarities),
                    'sample_count': len(similarities)
                }
                run_data['path_samples'][path_idx + 1] = high_reward_samples
                all_similarities.extend(similarities)
            else:
                run_data['path_similarities'][path_idx + 1] = {
                    'avg_similarity': 0.0,
                    'max_similarity': 0.0,
                    'min_similarity': 0.0,
                    'sample_count': 0
                }
                run_data['path_samples'][path_idx + 1] = []

        if all_similarities:
            run_data['overall_avg_similarity'] = np.mean(all_similarities)
            run_data['max_similarity'] = np.max(all_similarities)
            run_data['min_similarity'] = np.min(all_similarities)
        else:
            run_data['overall_avg_similarity'] = 0.0
            run_data['max_similarity'] = 0.0
            run_data['min_similarity'] = 0.0

        all_runs_data.append(run_data)

        print(f"[Run {run_id}] completed! Overall Average Similarity: {run_data['overall_avg_similarity']:.4f}")
        print(f"{'=' * 60}\n")

    total_time = time.time() - total_start_time

    print("\nGenerating consolidated Excel report...")
    create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir)

    print("\n" + "=" * 60)
    print(f"\nAll results have been saved to: {output_dir}")
    print("=" * 60)

if __name__ == "__main__":
    run_20_times_training()