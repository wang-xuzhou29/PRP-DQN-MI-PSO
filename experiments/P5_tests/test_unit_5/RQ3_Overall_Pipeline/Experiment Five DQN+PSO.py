import os
import sys
import random
import time
from collections import deque
from datetime import datetime
from typing import Any, Dict, List, Optional, Set, Tuple

import numpy as np
import torch
import torch.nn as nn
import torch.optim as optim
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ========================================
# ========== experiment configuration parameters ==========
# ========================================
NUM_RUNS = 20

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === three-dimensional range settings ===
# Keep the current DQN state range used by the second script. To use a 0-500 range, modify this section only.
LIGHT_MIN = 1
LIGHT_MAX = 100
MOISTURE_MIN = 1
MOISTURE_MAX = 400
TEMP_MIN = 1
TEMP_MAX = 1000
BOUNDS = {
    "light": (LIGHT_MIN, LIGHT_MAX),
    "moisture": (MOISTURE_MIN, MOISTURE_MAX),
    "temp": (TEMP_MIN, TEMP_MAX),
}

# === standard PSO parameters ===
PSO_W = 0.7
PSO_C1 = 1.5
PSO_C2 = 1.5
PSO_VMAX_RATIO = 0.2


def normalize_state(state):
    """Normalize the state to the [0, 1] range."""
    return np.array([
        (state[0] - BOUNDS["light"][0]) / (BOUNDS["light"][1] - BOUNDS["light"][0]),
        (state[1] - BOUNDS["moisture"][0]) / (BOUNDS["moisture"][1] - BOUNDS["moisture"][0]),
        (state[2] - BOUNDS["temp"][0]) / (BOUNDS["temp"][1] - BOUNDS["temp"][0]),
    ], dtype=np.float32)


def denormalize_state(normalized_state):
    """Restore the normalized state to the original range."""
    return np.array([
        normalized_state[0] * (BOUNDS["light"][1] - BOUNDS["light"][0]) + BOUNDS["light"][0],
        normalized_state[1] * (BOUNDS["moisture"][1] - BOUNDS["moisture"][0]) + BOUNDS["moisture"][0],
        normalized_state[2] * (BOUNDS["temp"][1] - BOUNDS["temp"][0]) + BOUNDS["temp"][0],
    ], dtype=float)


def generate_random_state():
    """Generate a random state within each dimension range."""
    light = np.random.randint(BOUNDS["light"][0], BOUNDS["light"][1] + 1)
    moisture = np.random.randint(BOUNDS["moisture"][0], BOUNDS["moisture"][1] + 1)
    temp = np.random.randint(BOUNDS["temp"][0], BOUNDS["temp"][1] + 1)
    return np.array([light, moisture, temp], dtype=float)


def clip_state(state):
    """Clip the state to each dimension boundary."""
    return np.array([
        np.clip(state[0], BOUNDS["light"][0], BOUNDS["light"][1]),
        np.clip(state[1], BOUNDS["moisture"][0], BOUNDS["moisture"][1]),
        np.clip(state[2], BOUNDS["temp"][0], BOUNDS["temp"][1]),
    ], dtype=float)


def is_state_valid(state):
    """Check whether the state is within all valid dimension ranges."""
    return (
        BOUNDS["light"][0] <= state[0] <= BOUNDS["light"][1]
        and BOUNDS["moisture"][0] <= state[1] <= BOUNDS["moisture"][1]
        and BOUNDS["temp"][0] <= state[2] <= BOUNDS["temp"][1]
    )


def execute_Tr(special_event, emergency_vehicles, visibility):
    triggered = set()

    if (special_event < 50) != (special_event < 70):
        triggered.add(1)
    if (special_event > 50) != (special_event > 30):
        triggered.add(2)
    if (special_event > 50) != (special_event > 70):
        triggered.add(3)
    if (special_event <= 25) != (special_event <= 75):
        triggered.add(4)
    if (special_event >= 25) != (special_event >= 75):
        triggered.add(5)
    if (special_event <= 50) != (special_event <= 25):
        triggered.add(6)

    if (emergency_vehicles < 200) != (emergency_vehicles < 100):
        triggered.add(7)
    if (emergency_vehicles < 200) != (emergency_vehicles < 300):
        triggered.add(8)
    if (emergency_vehicles > 200) != (emergency_vehicles > 100):
        triggered.add(9)
    if (emergency_vehicles > 200) != (emergency_vehicles > 300):
        triggered.add(10)
    if (emergency_vehicles <= 150) != (emergency_vehicles <= 250):
        triggered.add(11)
    if (emergency_vehicles >= 150) != (emergency_vehicles >= 250):
        triggered.add(12)
    if (emergency_vehicles % 2 == 0) != (emergency_vehicles % 3 == 0):
        triggered.add(13)

    if (visibility < 500) != (visibility < 300):
        triggered.add(14)
    if (visibility < 500) != (visibility < 700):
        triggered.add(15)
    if (visibility > 500) != (visibility > 300):
        triggered.add(16)
    if (visibility > 500) != (visibility > 700):
        triggered.add(17)
    if (visibility <= 400) != (visibility <= 600):
        triggered.add(18)
    if (visibility >= 400) != (visibility >= 600):
        triggered.add(19)
    if (visibility % 10 < 5) != (visibility % 20 < 5):
        triggered.add(20)

    if (special_event < 50 and emergency_vehicles < 200) != (
            special_event < 30 and emergency_vehicles < 200):
        triggered.add(21)
    if (special_event < 50 and emergency_vehicles < 200) != (
            special_event < 50 and emergency_vehicles < 150):
        triggered.add(22)
    if (special_event > 50 and emergency_vehicles > 200) != (
            special_event > 30 and emergency_vehicles > 200):
        triggered.add(23)
    if (special_event > 50 and emergency_vehicles > 200) != (
            special_event > 50 and emergency_vehicles > 150):
        triggered.add(24)
    if (special_event <= 40 or emergency_vehicles <= 180) != (
            special_event <= 60 or emergency_vehicles <= 180):
        triggered.add(25)
    if (special_event <= 40 or emergency_vehicles <= 180) != (
            special_event <= 40 or emergency_vehicles <= 320):
        triggered.add(26)

    if (special_event < 50 and visibility < 500) != (special_event < 30 and visibility < 500):
        triggered.add(27)
    if (special_event < 50 and visibility < 500) != (special_event < 50 and visibility < 700):
        triggered.add(28)
    if (special_event > 50 and visibility > 500) != (special_event > 30 and visibility > 500):
        triggered.add(29)
    if (special_event > 50 and visibility > 500) != (special_event > 50 and visibility > 200):
        triggered.add(30)
    if (special_event <= 60 or visibility <= 600) != (special_event <= 40 or visibility <= 600):
        triggered.add(31)
    if (special_event <= 60 or visibility <= 600) != (special_event <= 60 or visibility <= 400):
        triggered.add(32)

    if (emergency_vehicles < 200 and visibility < 500) != (
            emergency_vehicles < 150 and visibility < 500):
        triggered.add(33)
    if (emergency_vehicles < 200 and visibility < 500) != (
            emergency_vehicles < 200 and visibility < 800):
        triggered.add(34)
    if (emergency_vehicles > 200 and visibility > 500) != (
            emergency_vehicles > 150 and visibility > 500):
        triggered.add(35)
    if (emergency_vehicles > 200 and visibility > 500) != (
            emergency_vehicles > 200 and visibility > 200):
        triggered.add(36)
    if (emergency_vehicles <= 350 or visibility <= 700) != (
            emergency_vehicles <= 180 or visibility <= 700):
        triggered.add(37)
    if (emergency_vehicles <= 250 or visibility <= 700) != (
            emergency_vehicles <= 250 or visibility <= 500):
        triggered.add(38)

    if (special_event < 40 or emergency_vehicles > 300) != (
            special_event < 60 or emergency_vehicles > 300):
        triggered.add(39)
    if (special_event < 40 or emergency_vehicles > 300) != (
            special_event < 40 or emergency_vehicles > 250):
        triggered.add(40)
    if (special_event > 60 or visibility < 400) != (special_event > 40 or visibility < 400):
        triggered.add(41)
    if (special_event > 60 or visibility < 400) != (special_event > 60 or visibility < 300):
        triggered.add(42)
    if (emergency_vehicles < 150 or visibility > 700) != (emergency_vehicles < 200 or visibility > 700):
        triggered.add(43)
    if (emergency_vehicles < 150 or visibility > 700) != (emergency_vehicles < 150 or visibility > 400):
        triggered.add(44)

    if (special_event + emergency_vehicles < 350) != (special_event + emergency_vehicles < 200):
        triggered.add(45)
    if (special_event * 2 < emergency_vehicles) != (special_event * 3 < emergency_vehicles):
        triggered.add(46)
    if (visibility - special_event > 400) != (visibility - special_event > 500):
        triggered.add(47)
    if (emergency_vehicles % 50 == 0) != (emergency_vehicles % 100 == 0):
        triggered.add(48)
    if (special_event + visibility > 600) != (special_event + visibility > 500):
        triggered.add(49)
    if (abs(special_event - 50) < 20) != (abs(special_event - 50) < 30):
        triggered.add(50)

    if (special_event < 60 and emergency_vehicles < 250 and visibility > 400) != (
            special_event < 40 and emergency_vehicles < 250 and visibility > 400):
        triggered.add(51)
    if (special_event < 60 and emergency_vehicles < 250 and visibility > 400) != (
            special_event < 60 and emergency_vehicles < 200 and visibility > 400):
        triggered.add(52)
    if (special_event < 60 and emergency_vehicles < 250 and visibility > 400) != (
            special_event < 60 and emergency_vehicles < 250 and visibility > 300):
        triggered.add(53)

    if (special_event > 40 and emergency_vehicles > 150 and visibility < 600) != (
            special_event > 60 and emergency_vehicles > 150 and visibility < 600):
        triggered.add(54)
    if (special_event > 40 and emergency_vehicles > 150 and visibility < 600) != (
            special_event > 40 and emergency_vehicles > 200 and visibility < 600):
        triggered.add(55)
    if (special_event > 40 and emergency_vehicles > 150 and visibility < 600) != (
            special_event > 40 and emergency_vehicles > 150 and visibility < 500):
        triggered.add(56)

    if (special_event < 30 or emergency_vehicles < 100 or visibility > 800) != (
            special_event < 50 or emergency_vehicles < 100 or visibility > 800):
        triggered.add(57)
    if (special_event < 30 or emergency_vehicles < 100 or visibility > 800) != (
            special_event < 30 or emergency_vehicles < 150 or visibility > 800):
        triggered.add(58)
    if (special_event < 30 or emergency_vehicles < 100 or visibility > 800) != (
            special_event < 30 or emergency_vehicles < 100 or visibility > 700):
        triggered.add(59)

    if (special_event > 70 or emergency_vehicles > 300 or visibility < 200) != (
            special_event > 50 or emergency_vehicles > 300 or visibility < 200):
        triggered.add(60)
    if (special_event > 70 or emergency_vehicles > 300 or visibility < 200) != (
            special_event > 70 or emergency_vehicles > 250 or visibility < 200):
        triggered.add(61)
    if (special_event > 70 or emergency_vehicles > 300 or visibility < 200) != (
            special_event > 70 or emergency_vehicles > 300 or visibility < 300):
        triggered.add(62)

    if ((special_event < 50 and emergency_vehicles < 200) or visibility > 700) != (
            (special_event < 30 and emergency_vehicles < 200) or visibility > 700):
        triggered.add(63)
    if ((special_event < 50 and emergency_vehicles < 200) or visibility > 700) != (
            (special_event < 50 and emergency_vehicles < 150) or visibility > 700):
        triggered.add(64)
    if ((special_event < 50 and emergency_vehicles < 200) or visibility > 700) != (
            (special_event < 50 and emergency_vehicles < 200) or visibility > 400):
        triggered.add(65)

    if ((special_event > 50 or emergency_vehicles > 200) and visibility < 500) != (
            (special_event > 30 or emergency_vehicles > 200) and visibility < 500):
        triggered.add(66)
    if ((special_event > 50 or emergency_vehicles > 200) and visibility < 500) != (
            (special_event > 50 or emergency_vehicles > 150) and visibility < 500):
        triggered.add(67)
    if ((special_event > 50 or emergency_vehicles > 200) and visibility < 500) != (
            (special_event > 50 or emergency_vehicles > 200) and visibility < 400):
        triggered.add(68)

    if (special_event + emergency_vehicles > 300 and visibility < 600) != (
            special_event + emergency_vehicles > 250 and visibility < 600):
        triggered.add(69)
    if (special_event + emergency_vehicles > 300 and visibility < 600) != (
            special_event + emergency_vehicles > 300 and visibility < 400):
        triggered.add(70)

    if (special_event * 4 < visibility and emergency_vehicles > 100) != (
            special_event * 5 < visibility and emergency_vehicles > 100):
        triggered.add(71)
    if (special_event * 4 < visibility and emergency_vehicles > 100) != (
            special_event * 4 < visibility and emergency_vehicles > 150):
        triggered.add(72)

    if (visibility - emergency_vehicles > 200 and special_event > 30) != (
            visibility - emergency_vehicles > 300 and special_event > 30):
        triggered.add(73)
    if (visibility - emergency_vehicles > 200 and special_event > 30) != (
            visibility - emergency_vehicles > 200 and special_event > 50):
        triggered.add(74)

    if (special_event % 3 == 0 and emergency_vehicles > 120 and visibility <= 750) != (
            special_event % 4 == 0 and emergency_vehicles > 120 and visibility <= 750):
        triggered.add(75)
    if (special_event % 3 == 0 and emergency_vehicles > 120 and visibility <= 750) != (
            special_event % 3 == 0 and emergency_vehicles > 180 and visibility <= 750):
        triggered.add(76)
    if (special_event % 3 == 0 and emergency_vehicles > 120 and visibility <= 750) != (
            special_event % 3 == 0 and emergency_vehicles > 120 and visibility <= 650):
        triggered.add(77)

    if (abs(special_event - 50) > 10 and emergency_vehicles < 300 and visibility % 100 < 50) != (
            abs(special_event - 50) > 20 and emergency_vehicles < 300 and visibility % 100 < 50):
        triggered.add(78)
    if (abs(special_event - 50) > 10 and emergency_vehicles < 300 and visibility % 100 < 50) != (
            abs(special_event - 50) > 10 and emergency_vehicles < 250 and visibility % 100 < 50):
        triggered.add(79)
    if (abs(special_event - 50) > 10 and emergency_vehicles < 300 and visibility % 100 < 50) != (
            abs(special_event - 50) > 10 and emergency_vehicles < 300 and visibility % 100 < 30):
        triggered.add(80)

    if (20 <= special_event <= 80 and emergency_vehicles < 250) != (
            30 <= special_event <= 80 and emergency_vehicles < 250):
        triggered.add(81)
    if (20 <= special_event <= 80 and emergency_vehicles < 250) != (
            20 <= special_event <= 70 and emergency_vehicles < 250):
        triggered.add(82)
    if (20 <= special_event <= 80 and emergency_vehicles < 250) != (
            20 <= special_event <= 80 and emergency_vehicles < 200):
        triggered.add(83)

    if (100 <= emergency_vehicles <= 300 and visibility > 400) != (
            150 <= emergency_vehicles <= 300 and visibility > 400):
        triggered.add(84)
    if (100 <= emergency_vehicles <= 300 and visibility > 400) != (
            100 <= emergency_vehicles <= 250 and visibility > 400):
        triggered.add(85)
    if (100 <= emergency_vehicles <= 300 and visibility > 400) != (
            100 <= emergency_vehicles <= 300 and visibility > 300):
        triggered.add(86)

    if (300 <= visibility <= 700 and special_event > 25) != (
            400 <= visibility <= 700 and special_event > 25):
        triggered.add(87)
    if (300 <= visibility <= 700 and special_event > 25) != (
            300 <= visibility <= 600 and special_event > 25):
        triggered.add(88)
    if (300 <= visibility <= 700 and special_event > 25) != (
            300 <= visibility <= 700 and special_event > 35):
        triggered.add(89)

    if ((10 <= special_event <= 40) and (150 <= emergency_vehicles <= 350)) != (
            (20 <= special_event <= 40) and (150 <= emergency_vehicles <= 350)):
        triggered.add(90)
    if ((10 <= special_event <= 40) and (150 <= emergency_vehicles <= 350)) != (
            (10 <= special_event <= 30) and (150 <= emergency_vehicles <= 350)):
        triggered.add(91)
    if ((10 <= special_event <= 40) and (150 <= emergency_vehicles <= 350)) != (
            (10 <= special_event <= 40) and (200 <= emergency_vehicles <= 350)):
        triggered.add(92)
    if ((10 <= special_event <= 40) and (150 <= emergency_vehicles <= 350)) != (
            (10 <= special_event <= 40) and (150 <= emergency_vehicles <= 300)):
        triggered.add(93)

    if ((special_event > 60) and (200 < visibility < 800)) != (
            (special_event > 50) and (200 < visibility < 800)):
        triggered.add(94)
    if ((special_event > 60) and (200 < visibility < 800)) != (
            (special_event > 60) and (300 < visibility < 800)):
        triggered.add(95)
    if ((special_event > 60) and (200 < visibility < 800)) != (
            (special_event > 60) and (200 < visibility < 700)):
        triggered.add(96)

    if (special_event < 45 and (emergency_vehicles > 180 or visibility < 450)) != (
            special_event < 55 and (emergency_vehicles > 180 or visibility < 450)):
        triggered.add(97)
    if (special_event < 45 and (emergency_vehicles > 180 or visibility < 450)) != (
            special_event < 45 and (emergency_vehicles > 220 or visibility < 450)):
        triggered.add(98)
    if (special_event < 45 and (emergency_vehicles > 180 or visibility < 450)) != (
            special_event < 45 and (emergency_vehicles > 180 or visibility < 350)):
        triggered.add(99)

    if (emergency_vehicles > 175 and (special_event < 35 or visibility > 650)) != (
            emergency_vehicles > 225 and (special_event < 35 or visibility > 650)):
        triggered.add(100)
    if (emergency_vehicles > 175 and (special_event < 35 or visibility > 650)) != (
            emergency_vehicles > 175 and (special_event < 45 or visibility > 650)):
        triggered.add(101)
    if (emergency_vehicles > 175 and (special_event < 35 or visibility > 650)) != (
            emergency_vehicles > 175 and (special_event < 35 or visibility > 550)):
        triggered.add(102)

    if (special_event + emergency_vehicles // 2 > 150) != (
            special_event + emergency_vehicles // 3 > 150):
        triggered.add(103)
    if (visibility // 10 + special_event > 80) != (visibility // 15 + special_event > 80):
        triggered.add(104)
    if (emergency_vehicles % 10 + special_event % 10 < 15) != (
            emergency_vehicles % 15 + special_event % 10 < 15):
        triggered.add(105)
    if (emergency_vehicles % 10 + special_event % 10 < 15) != (
            emergency_vehicles % 10 + special_event % 15 < 15):
        triggered.add(106)

    if (max(special_event, emergency_vehicles // 4) > 45) != (
            max(special_event, emergency_vehicles // 5) > 45):
        triggered.add(107)
    if (min(visibility // 10, special_event) < 35) != (min(visibility // 15, special_event) < 35):
        triggered.add(108)
    if (min(visibility // 10, special_event) < 35) != (min(visibility // 10, special_event) < 25):
        triggered.add(109)
    if (special_event > visibility // 20) != (special_event > visibility // 25):
        triggered.add(110)

    if (special_event % 2 == emergency_vehicles % 2) != (special_event % 3 == emergency_vehicles % 2):
        triggered.add(111)
    if (special_event % 2 == emergency_vehicles % 2) != (special_event % 2 == emergency_vehicles % 3):
        triggered.add(112)
    if (visibility % 7 == 0) != (visibility % 11 == 0):
        triggered.add(113)
    if ((special_event + emergency_vehicles) % 5 == 0) != (
            (special_event + emergency_vehicles) % 7 == 0):
        triggered.add(114)

    if (special_event & 1 == visibility & 1) != (special_event & 3 == visibility & 1):
        triggered.add(115)
    if (special_event & 1 == visibility & 1) != (special_event & 1 == visibility & 3):
        triggered.add(116)
    if (emergency_vehicles >> 2 > special_event) != (emergency_vehicles >> 3 > special_event):
        triggered.add(117)
    if ((special_event ^ emergency_vehicles) % 50 < 25) != (
            (special_event ^ emergency_vehicles) % 60 < 25):
        triggered.add(118)
    if ((special_event ^ emergency_vehicles) % 50 < 25) != (
            (special_event ^ emergency_vehicles) % 50 < 30):
        triggered.add(119)

    if (len(str(special_event)) + len(str(emergency_vehicles)) > 4) != (
            len(str(special_event)) + len(str(emergency_vehicles)) > 5):
        triggered.add(120)
    if (sum(int(d) for d in str(visibility)) > 10) != (sum(int(d) for d in str(visibility)) > 12):
        triggered.add(121)
    if (special_event ** 2 % 100 < 50) != (special_event ** 2 % 120 < 50):
        triggered.add(122)
    if (special_event ** 2 % 100 < 50) != (special_event ** 2 % 100 < 60):
        triggered.add(123)

    if (abs(special_event - emergency_vehicles // 4) < 20) != (
            abs(special_event - emergency_vehicles // 5) < 20):
        triggered.add(124)
    if (abs(special_event - emergency_vehicles // 4) < 20) != (
            abs(special_event - emergency_vehicles // 4) < 25):
        triggered.add(125)
    if ((visibility + special_event + emergency_vehicles) % 100 > 50) != (
            (visibility + special_event + emergency_vehicles) % 120 > 50):
        triggered.add(126)

    return triggered



target_paths = [
    [2, 4, 5, 6, 8, 10, 11, 12, 15, 17, 18, 19, 20, 23, 28, 29, 44, 45, 49, 52, 57, 65, 74, 83, 89, 91, 98, 101, 102, 104, 105, 107, 111, 112, 113, 115, 117, 118, 120],
    [2, 4, 5, 6, 7, 9, 11, 12, 13, 14, 16, 18, 19, 21, 22, 27, 33, 43, 44, 45, 47, 57, 63, 64, 66, 67, 73, 74, 76, 91, 92, 104, 108, 112, 116, 117, 118, 120, 121],
    [2, 4, 5, 6, 7, 9, 11, 12, 13, 14, 16, 20, 21, 22, 27, 33, 42, 43, 45, 53, 57, 63, 64, 66, 67, 75, 86, 87, 89, 91, 92, 100, 109, 112, 114, 117, 118, 120, 122],
    [2, 4, 5, 6, 7, 9, 11, 12, 13, 14, 16, 18, 19, 21, 22, 27, 33, 43, 44, 47, 49, 57, 63, 64, 66, 67, 74, 76, 91, 92, 104, 108, 112, 114, 117, 118, 120],
    [7, 9, 11, 12, 13, 15, 17, 18, 19, 20, 22, 28, 34, 35, 43, 44, 45, 47, 49, 50, 64, 75, 81, 92, 98, 100, 110, 111, 112, 118, 120, 122, 124, 125, 126],
    [2, 4, 5, 6, 8, 10, 11, 12, 20, 23, 25, 26, 29, 31, 37, 39, 41, 45, 51, 52, 74, 83, 100, 103, 105, 106, 107, 110, 113, 114, 115, 117, 118, 120],
    [2, 4, 5, 6, 8, 9, 11, 12, 13, 15, 16, 18, 19, 28, 44, 45, 47, 49, 52, 57, 65, 73, 74, 78, 83, 91, 98, 101, 104, 107, 108, 112, 117, 120],
    [1, 3, 4, 5, 8, 10, 13, 14, 16, 18, 19, 26, 30, 32, 36, 40, 44, 45, 49, 60, 61, 65, 68, 70, 78, 79, 80, 85, 108, 111, 112, 116, 117, 120],
    [2, 4, 5, 6, 8, 10, 11, 12, 14, 16, 23, 25, 26, 27, 36, 39, 42, 45, 53, 54, 57, 69, 75, 83, 86, 87, 97, 103, 109, 111, 117, 118, 120, 122],
    [1, 3, 4, 5, 8, 10, 11, 12, 14, 16, 25, 26, 30, 36, 39, 42, 45, 53, 54, 60, 69, 75, 83, 86, 87, 94, 103, 104, 109, 111, 113, 117, 120, 121],
    [2, 4, 5, 6, 8, 10, 11, 12, 13, 20, 23, 25, 26, 29, 31, 37, 39, 41, 45, 51, 52, 57, 59, 74, 83, 97, 103, 111, 112, 113, 117, 118, 120, 122],
    [2, 4, 5, 6, 7, 9, 12, 15, 17, 18, 19, 20, 21, 22, 28, 29, 34, 43, 44, 48, 57, 63, 64, 72, 74, 76, 91, 92, 104, 116, 117, 119, 120],
    [1, 3, 4, 5, 8, 10, 11, 12, 14, 16, 25, 26, 30, 36, 39, 42, 45, 53, 54, 60, 69, 75, 83, 86, 87, 94, 97, 103, 104, 105, 106, 108, 120],
    [7, 9, 11, 12, 13, 15, 17, 22, 28, 34, 35, 43, 44, 45, 50, 64, 75, 77, 81, 92, 98, 100, 104, 106, 107, 111, 112, 116, 118, 120, 122, 124, 126],
    [8, 10, 11, 12, 14, 16, 18, 19, 26, 30, 32, 36, 44, 45, 49, 50, 65, 68, 70, 73, 75, 80, 82, 83, 108, 111, 113, 118, 120, 122, 124, 126],
    [4, 5, 6, 7, 9, 11, 12, 13, 14, 16, 18, 19, 22, 33, 43, 44, 50, 64, 67, 81, 89, 92, 99, 109, 112, 113, 115, 117, 119, 120],
    [1, 3, 4, 5, 8, 10, 11, 12, 13, 25, 26, 30, 36, 39, 45, 54, 60, 62, 69, 75, 83, 94, 103, 104, 105, 106, 109, 112, 118, 120],
    [1, 3, 4, 5, 13, 15, 17, 18, 19, 25, 38, 41, 44, 47, 49, 54, 56, 65, 70, 75, 94, 97, 108, 112, 113, 115, 116, 117, 120, 121],
    [7, 9, 11, 12, 14, 16, 18, 19, 24, 30, 32, 33, 43, 44, 45, 49, 55, 65, 68, 69, 71, 73, 75, 80, 103, 108, 111, 113, 118],
    [7, 9, 11, 12, 13, 15, 17, 22, 28, 34, 35, 43, 44, 64, 75, 90, 92, 104, 105, 111, 112, 119, 120, 122, 123, 124, 125, 126],
    [7, 9, 11, 12, 13, 20, 24, 26, 30, 33, 43, 45, 46, 55, 69, 75, 80, 95, 103, 106, 111, 112, 114, 120, 122],
    [7, 9, 11, 12, 24, 34, 35, 45, 46, 50, 59, 75, 76, 77, 82, 96, 103, 105, 106, 113, 115, 120, 121],
    [2, 4, 5, 6, 7, 9, 21, 29, 31, 39, 41, 51, 72, 74, 84, 106, 110, 111, 113, 114, 118, 120, 122],
    [7, 9, 13, 15, 17, 34, 45, 58, 65, 72, 80, 84, 88, 111, 112, 120, 126],
    [4, 5, 6, 13, 37, 48, 50, 93, 103, 109, 112, 114, 120, 122]
]


def jaccard_similarity(set1: Set[int], set2: Set[int]) -> float:
    """If set1 covers target set2, the match score is treated as 1."""
    if set2.issubset(set1):
        return 1.0
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    return intersection / union if union != 0 else 0.0


def compute_reward(state, target_path, triggered):
    """Compute the DQN reward."""
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward


class SimpleReplayBuffer:
    def __init__(self, capacity=5000):
        self.buffer = deque(maxlen=capacity)

    def append(self, experience):
        self.buffer.append(experience)

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return random.sample(self.buffer, len(self.buffer))
        return random.sample(self.buffer, batch_size)

    def __len__(self):
        return len(self.buffer)


class SimpleDQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super().__init__()
        self.fc1 = nn.Linear(state_dim, 64)
        self.fc2 = nn.Linear(64, 32)
        self.fc3 = nn.Linear(32, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)


class SimpleDQNAgent:
    def __init__(self, state_dim, action_dim, learning_rate=0.001):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.epsilon = 0.3
        self.gamma = 0.99

        self.model = SimpleDQN(state_dim, action_dim).to(device)
        self.target_model = SimpleDQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())
        self.replay_buffer = SimpleReplayBuffer(capacity=5000)

    def decode_action(self, action_idx):
        """Decode the action."""
        light_deltas = [int((LIGHT_MAX - LIGHT_MIN) * p) for p in [0.05, 0.1, 0.2, 0.5, 0.7, -0.05, -0.1, -0.2, -0.5, -0.7]]
        moisture_deltas = [int((MOISTURE_MAX - MOISTURE_MIN) * p) for p in [0.05, 0.1, 0.2, 0.5, 0.7, -0.05, -0.1, -0.2, -0.5, -0.7]]
        temp_deltas = [int((TEMP_MAX - TEMP_MIN) * p) for p in [0.05, 0.1, 0.2, 0.5, 0.7, -0.05, -0.1, -0.2, -0.5, -0.7]]

        dim = action_idx // 10
        delta_idx = action_idx % 10

        if dim == 0:
            return (light_deltas[delta_idx], 0, 0)
        if dim == 1:
            return (0, moisture_deltas[delta_idx], 0)
        return (0, 0, temp_deltas[delta_idx])

    def store_transition(self, state, action, reward, next_state, done):
        self.replay_buffer.append((state, action, reward, next_state, done))

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return

        batch = self.replay_buffer.sample(batch_size)
        states, actions, rewards, next_states, dones = zip(*batch)

        states = torch.tensor(np.array(states), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array(next_states), dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))

        loss = nn.MSELoss()(current_q_values, target_q_values)

        self.optimizer.zero_grad()
        loss.backward()
        self.optimizer.step()

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())

    def get_best_samples(self, target_path, num_samples=20):
        """Screen DQN candidate samples from the replay buffer. This stage is counted separately in T_sample_screening."""
        if len(self.replay_buffer) == 0:
            return []

        samples_with_scores = []
        for state_norm, _, _, _, _ in self.replay_buffer.buffer:
            state_original = denormalize_state(state_norm)
            triggered = execute_Tr(state_original)
            reward = compute_reward(state_original, target_path, triggered)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_scores.append((state_original, reward, sim, triggered))

        samples_with_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_scores[:num_samples]


def train_dqn_for_path(path_idx, target_path, num_samples=200):
    """Train a standard DQN for a single path."""
    print(f"  Start training path {path_idx + 1} DQN model...")
    agent = SimpleDQNAgent(state_dim=3, action_dim=30)

    random_states = [normalize_state(generate_random_state()) for _ in range(num_samples)]

    STEPS_PER_SAMPLE = 3
    EPOCHS = 5
    BATCH_SIZE = 32

    step_count = 0
    for _ in range(EPOCHS):
        for state_norm in random_states:
            state_norm = tuple(state_norm)
            state_original = denormalize_state(state_norm)

            for step in range(STEPS_PER_SAMPLE):
                legal_actions = []
                for a in range(agent.action_dim):
                    dx, dy, dz = agent.decode_action(a)
                    next_state_candidate = (state_original[0] + dx, state_original[1] + dy, state_original[2] + dz)
                    if is_state_valid(next_state_candidate):
                        legal_actions.append(a)

                if not legal_actions:
                    break

                if random.random() < agent.epsilon:
                    action = random.choice(legal_actions)
                else:
                    state_tensor = torch.tensor(state_norm, dtype=torch.float32).unsqueeze(0).to(device)
                    with torch.no_grad():
                        q_values = agent.model(state_tensor)[0]
                    action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                dx, dy, dz = agent.decode_action(action)
                next_state_original = (state_original[0] + dx, state_original[1] + dy, state_original[2] + dz)
                next_state_norm = normalize_state(next_state_original)

                triggered = execute_Tr(next_state_original)
                reward = compute_reward(next_state_original, target_path, triggered)
                done = (step == STEPS_PER_SAMPLE - 1)

                agent.store_transition(state_norm, action, reward, next_state_norm, done)

                state_norm = next_state_norm
                state_original = next_state_original
                step_count += 1

                if step_count % 50 == 0 and len(agent.replay_buffer) >= BATCH_SIZE:
                    agent.train(BATCH_SIZE)

                if step_count % 100 == 0:
                    agent.update_target_model()

    print(f"  Path {path_idx + 1} training completed, replay buffer size: {len(agent.replay_buffer)}")
    return agent


class Particle:
    def __init__(self, initial_position=None):
        if initial_position is not None:
            self.position = clip_state(np.array(initial_position, dtype=float))
        else:
            self.position = np.array([
                np.random.uniform(BOUNDS["light"][0], BOUNDS["light"][1]),
                np.random.uniform(BOUNDS["moisture"][0], BOUNDS["moisture"][1]),
                np.random.uniform(BOUNDS["temp"][0], BOUNDS["temp"][1]),
            ], dtype=float)

        spans = np.array([
            BOUNDS["light"][1] - BOUNDS["light"][0],
            BOUNDS["moisture"][1] - BOUNDS["moisture"][0],
            BOUNDS["temp"][1] - BOUNDS["temp"][0],
        ], dtype=float)
        self.velocity = np.random.uniform(-0.1 * spans, 0.1 * spans)
        self.best_position = self.position.copy()
        self.best_fitness = -1.0
        self.fitness = -1.0


class BasicPSO:
    """
    Standard PSO: contains only standard velocity and position updates.
    Mutation coefficients, opposite particles, local-best reset, and other improved PSO mechanisms have been removed.
    """

    def __init__(self, target_path, swarm_size=20, dqn_samples=None):
        self.target_path = target_path
        self.swarm_size = swarm_size
        self.particles = []
        self.global_best_position = None
        self.global_best_fitness = -1.0

        if dqn_samples is not None and len(dqn_samples) > 0:
            num_direct = min(len(dqn_samples), swarm_size)
            for i in range(num_direct):
                state_tuple, _, _, _ = dqn_samples[i]
                self.particles.append(Particle(initial_position=state_tuple))

            while len(self.particles) < swarm_size:
                base_idx = len(self.particles) % len(dqn_samples)
                state_tuple, _, _, _ = dqn_samples[base_idx]
                perturbed = np.array(state_tuple, dtype=float) + np.random.randint(-10, 11, size=3)
                self.particles.append(Particle(initial_position=clip_state(perturbed)))
        else:
            self.particles = [Particle() for _ in range(swarm_size)]

        for particle in self.particles:
            particle.fitness = self.fitness_function(particle.position)
            particle.best_fitness = particle.fitness
            particle.best_position = particle.position.copy()

            if particle.fitness > self.global_best_fitness:
                self.global_best_fitness = particle.fitness
                self.global_best_position = particle.position.copy()

    def fitness_function(self, position):
        try:
            triggered = execute_Tr(position)
            return jaccard_similarity(triggered, self.target_path)
        except Exception:
            return 0.0

    def update(self):
        spans = np.array([
            BOUNDS["light"][1] - BOUNDS["light"][0],
            BOUNDS["moisture"][1] - BOUNDS["moisture"][0],
            BOUNDS["temp"][1] - BOUNDS["temp"][0],
        ], dtype=float)
        max_velocity = PSO_VMAX_RATIO * spans

        for particle in self.particles:
            r1 = np.random.random(3)
            r2 = np.random.random(3)

            particle.velocity = (
                PSO_W * particle.velocity
                + PSO_C1 * r1 * (particle.best_position - particle.position)
                + PSO_C2 * r2 * (self.global_best_position - particle.position)
            )
            particle.velocity = np.clip(particle.velocity, -max_velocity, max_velocity)

            particle.position = clip_state(particle.position + particle.velocity)
            particle.fitness = self.fitness_function(particle.position)

            if particle.fitness > particle.best_fitness:
                particle.best_fitness = particle.fitness
                particle.best_position = particle.position.copy()

            if particle.fitness > self.global_best_fitness:
                self.global_best_fitness = particle.fitness
                self.global_best_position = particle.position.copy()


def _set_header(ws, headers, widths, header_fill, header_font, border, center_align):
    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_row(ws, row_idx, row_data, border, alignment, fill=None):
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.border = border
        cell.alignment = alignment
        if fill is not None:
            cell.fill = fill


def export_time_metrics_to_excel(all_run_results, experiment_total_time=None, filename=None):
    """Export only timing metrics and per-path iteration counts."""
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"DQN_BasicPSO_Time_Metrics_{timestamp}.xlsx"

    output_dir = os.path.join(os.getcwd(), "results")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    num_runs = len(all_run_results)
    num_paths = len(target_paths)

    # Worksheet 1: runtime summary
    ws1 = wb.active
    ws1.title = "Runtime Summary"
    ws1.sheet_view.showGridLines = False
    headers1 = [
        "Run", "Number of Paths", "DQNTraining Total Time(s)", "Sample Screening Total Time(s)", "DQNDirect Check Total Time(s)",
        "PSOInitialization Total Time(s)", "PSOSearch Total Time(s)", "Total Algorithm Time(s)", "Total Iterations", "Average Iterations per Path"
    ]
    widths1 = [12, 10, 18, 18, 22, 18, 18, 16, 14, 18]
    _set_header(ws1, headers1, widths1, header_fill, header_font, border, center_align)

    for run_idx, run_result in enumerate(all_run_results, 1):
        path_results = run_result["paths"]
        total_dqn_train = sum(p["T_DQN_train"] for p in path_results)
        total_screen = sum(p["T_sample_screening"] for p in path_results)
        total_direct_check = sum(p["T_direct_check"] for p in path_results)
        total_pso_init = sum(p["T_PSO_init"] for p in path_results)
        total_pso_search = sum(p["T_PSO_search"] for p in path_results)
        total_iterations = sum(p["iterations"] for p in path_results)

        row_data = [
            f" {run_idx}",
            num_paths,
            round(total_dqn_train, 6),
            round(total_screen, 6),
            round(total_direct_check, 6),
            round(total_pso_init, 6),
            round(total_pso_search, 6),
            round(run_result["T_run_algorithm_total"], 6),
            int(total_iterations),
            round(total_iterations / num_paths, 2),
        ]
        _write_row(ws1, run_idx + 1, row_data, border, center_align, alternate_fill if run_idx % 2 == 0 else None)

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:J{num_runs + 1}"

    # Worksheet 2: path timing and iteration details
    ws2 = wb.create_sheet(title="Path Timing and Iteration Details")
    ws2.sheet_view.showGridLines = False
    headers2 = [
        "Run", "Path ID", "DQN training(s)", "Sample Screening Time(s)", "DQNDirect Check Time(s)",
        "PSOInitialization Time(s)", "PSOSearch Time(s)", "Path Total Algorithm Time(s)", "Iterations"
    ]
    widths2 = [12, 12, 18, 18, 22, 18, 18, 18, 12]
    _set_header(ws2, headers2, widths2, header_fill, header_font, border, center_align)

    row_idx = 2
    for run_idx, run_result in enumerate(all_run_results, 1):
        for path_result in run_result["paths"]:
            row_data = [
                f" {run_idx}",
                f"Path  {path_result['path_idx'] + 1}",
                round(path_result["T_DQN_train"], 6),
                round(path_result["T_sample_screening"], 6),
                round(path_result["T_direct_check"], 6),
                round(path_result["T_PSO_init"], 6),
                round(path_result["T_PSO_search"], 6),
                round(path_result["T_path_algorithm_total"], 6),
                int(path_result["iterations"]),
            ]
            _write_row(ws2, row_idx, row_data, border, center_align, alternate_fill if row_idx % 2 == 0 else None)
            row_idx += 1

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:I{row_idx - 1}"

    # Worksheet 3: path aggregate statistics
    ws3 = wb.create_sheet(title="Path Aggregate Statistics")
    ws3.sheet_view.showGridLines = False
    headers3 = [
        "Path ID", "Average DQN Training Time(s)", "Average Sample Screening Time(s)", "Average PSO Initialization Time(s)",
        "Average PSO Search Time(s)", "Average Path Total Algorithm Time(s)", "Average Iterations", "Minimum Iterations", "Maximum Iterations"
    ]
    widths3 = [12, 22, 22, 22, 22, 22, 16, 14, 14]
    _set_header(ws3, headers3, widths3, header_fill, header_font, border, center_align)

    for path_idx in range(num_paths):
        records = [run_result["paths"][path_idx] for run_result in all_run_results]
        iterations = [r["iterations"] for r in records]
        row_data = [
            f"Path  {path_idx + 1}",
            round(float(np.mean([r["T_DQN_train"] for r in records])), 6),
            round(float(np.mean([r["T_sample_screening"] for r in records])), 6),
            round(float(np.mean([r["T_PSO_init"] for r in records])), 6),
            round(float(np.mean([r["T_PSO_search"] for r in records])), 6),
            round(float(np.mean([r["T_path_algorithm_total"] for r in records])), 6),
            round(float(np.mean(iterations)), 2),
            int(np.min(iterations)),
            int(np.max(iterations)),
        ]
        _write_row(ws3, path_idx + 2, row_data, border, center_align, alternate_fill if (path_idx + 1) % 2 == 0 else None)

    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:I{num_paths + 1}"

    # Worksheet 4: experiment overview
    ws4 = wb.create_sheet(title="Experiment Overview")
    ws4.sheet_view.showGridLines = False
    headers4 = ["Metric", "Value"]
    widths4 = [32, 24]
    _set_header(ws4, headers4, widths4, header_fill, header_font, border, center_align)

    all_path_records = [p for run_result in all_run_results for p in run_result["paths"]]
    all_iterations = [p["iterations"] for p in all_path_records]
    overview_rows = [
        ["Experiment Runs", num_runs],
        ["Number of Target Paths", num_paths],
        ["Total Experiment Wall-clock Time(s)", round(experiment_total_time, 6) if experiment_total_time is not None else ""],
        ["Average Total Algorithm Time per Run(s)", round(float(np.mean([r["T_run_algorithm_total"] for r in all_run_results])), 6)],
        ["Average Total Algorithm Time per Path(s)", round(float(np.mean([p["T_path_algorithm_total"] for p in all_path_records])), 6)],
        ["Average DQN Training Time per Path(s)", round(float(np.mean([p["T_DQN_train"] for p in all_path_records])), 6)],
        ["Average Sample Screening Time per Path(s)", round(float(np.mean([p["T_sample_screening"] for p in all_path_records])), 6)],
        ["Average PSO Search Time per Path(s)", round(float(np.mean([p["T_PSO_search"] for p in all_path_records])), 6)],
        ["Average Iterations per Path", round(float(np.mean(all_iterations)), 2)],
        ["Maximum Iterations per Path", int(np.max(all_iterations))],
        ["Minimum Iterations per Path", int(np.min(all_iterations))],
    ]

    for idx, row_data in enumerate(overview_rows, 2):
        _write_row(ws4, idx, row_data, border, center_align, alternate_fill if idx % 2 == 0 else None)

    wb.save(filepath)
    print(f"\n Timing metrics exported to: {filepath}")
    return filepath


def run_single_experiment(run_num, max_iterations=3000):
    """Run one DQN + standard PSO ."""
    print(f"\n{'=' * 80}")
    print(f"Start run  {run_num}  run: DQN + standard PSO ")
    print(f"{'=' * 80}")

    run_start = time.perf_counter()
    path_results = []

    for path_idx, target_path in enumerate(target_paths):
        print(f"\nRun {run_num} run - Path {path_idx + 1}")
        path_total_start = time.perf_counter()

        # 1) DQN training
        dqn_train_start = time.perf_counter()
        agent = train_dqn_for_path(path_idx, target_path, num_samples=200)
        T_DQN_train = time.perf_counter() - dqn_train_start

        # 2) DQNSample Screening Time
        screening_start = time.perf_counter()
        dqn_samples = agent.get_best_samples(target_path, num_samples=20)
        T_sample_screening = time.perf_counter() - screening_start

        # 3) DQN
        direct_check_start = time.perf_counter()
        direct_solution_found = False
        for state_tuple, reward, sim, triggered in dqn_samples:
            if sim >= 1.0 or target_path.issubset(triggered):
                direct_solution_found = True
                break
        T_direct_check = time.perf_counter() - direct_check_start

        T_PSO_init = 0.0
        T_PSO_search = 0.0
        iterations_used = 0

        if direct_solution_found:
            print(
                f"  Path {path_idx + 1}: DQN | "
                f"DQN training {T_DQN_train:.6f}s | screening {T_sample_screening:.6f}s | iterations {iterations_used}"
            )
        else:
            # 4) standard PSOInitialization Time
            pso_init_start = time.perf_counter()
            pso = BasicPSO(target_path, swarm_size=20, dqn_samples=dqn_samples)
            T_PSO_init = time.perf_counter() - pso_init_start

            # 5) standard PSOSearch Time
            pso_search_start = time.perf_counter()
            iterations_used = max_iterations
            for iteration in range(max_iterations):
                pso.update()
                if pso.global_best_fitness >= 1.0:
                    iterations_used = iteration + 1
                    break
            T_PSO_search = time.perf_counter() - pso_search_start

            print(
                f"  Path {path_idx + 1}: "
                f"DQN training {T_DQN_train:.6f}s | screening {T_sample_screening:.6f}s | "
                f"PSO {T_PSO_init:.6f}s | PSO {T_PSO_search:.6f}s | "
                f"iterations {iterations_used}"
            )

        T_path_algorithm_total = time.perf_counter() - path_total_start

        path_results.append({
            "path_idx": path_idx,
            "T_DQN_train": T_DQN_train,
            "T_sample_screening": T_sample_screening,
            "T_direct_check": T_direct_check,
            "T_PSO_init": T_PSO_init,
            "T_PSO_search": T_PSO_search,
            "T_path_algorithm_total": T_path_algorithm_total,
            "iterations": iterations_used,
        })

    T_run_algorithm_total = time.perf_counter() - run_start
    print(f"\nRun {run_num} runcompleted | Total Algorithm Time {T_run_algorithm_total:.6f}s")
    return {
        "run_num": run_num,
        "paths": path_results,
        "T_run_algorithm_total": T_run_algorithm_total,
    }


def run_multiple_experiments(num_runs):
    print("\n" + "=" * 80)
    print(f"DQN + standard PSO - {num_runs}: Metric")
    print("=" * 80)
    print(f"Number of Target Paths: {len(target_paths)}")
    print("Statistics: DQN training, Sample Screening Time, PSOInitialization Time, PSOSearch Time, Path Iterations")
    print("=" * 80)

    all_run_results = []
    experiment_start = time.perf_counter()

    for run_num in range(1, num_runs + 1):
        run_result = run_single_experiment(run_num)
        all_run_results.append(run_result)

    experiment_total_time = time.perf_counter() - experiment_start
    print(f"\n{'=' * 80}")
    print(f"All {num_runs} runcompleted | Total Experiment Wall-clock Time {experiment_total_time:.6f}s")
    print(f"{'=' * 80}\n")
    return all_run_results, experiment_total_time


if __name__ == "__main__":
    print("=" * 80)
    print("DQN + standard PSO - MetricPath Iterations")
    print("=" * 80)
    print(f"Current configuration: Run = {NUM_RUNS}")
    print(f"Number of Paths: {len(target_paths)}")
    print(f"Device: {device}")
    print("=" * 80)

    if len(sys.argv) > 1:
        try:
            NUM_RUNS = int(sys.argv[1])
            print(f"Read from command line: Run = {NUM_RUNS}")
        except ValueError:
            print(f"Invalid command-line argument, using default number of runs {NUM_RUNS}")

    all_results, total_time = run_multiple_experiments(num_runs=NUM_RUNS)
    export_time_metrics_to_excel(all_results, total_time)
    print("\nProgram completed")
