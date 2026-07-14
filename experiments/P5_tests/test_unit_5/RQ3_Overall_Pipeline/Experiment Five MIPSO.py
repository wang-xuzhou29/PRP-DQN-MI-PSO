import os
import random
import numpy as np
import time
import psutil
import math
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

NUM_RUNS = 20  # 默认实验运行次数

# === 新的三维范围设置 ===
LIGHT_MIN = 1
LIGHT_MAX = 100
MOISTURE_MIN = 1
MOISTURE_MAX = 400
TEMP_MIN = 1
TEMP_MAX = 1000

BOUNDS = {
    'light': (LIGHT_MIN, LIGHT_MAX),
    'temp': (TEMP_MIN, TEMP_MAX),
    'moisture': (MOISTURE_MIN, MOISTURE_MAX)
}


def clip_state(state):
    """将状态限制在各维度的边界内"""
    return np.array([
        np.clip(state[0], BOUNDS['light'][0], BOUNDS['light'][1]),
        np.clip(state[1], BOUNDS['temp'][0], BOUNDS['temp'][1]),
        np.clip(state[2], BOUNDS['moisture'][0], BOUNDS['moisture'][1])
    ])


def execute_Tr(position):
    """执行目标函数并返回触发的路径"""
    x = int(np.clip(position[0], BOUNDS['light'][0], BOUNDS['light'][1]))
    temp = int(np.clip(position[1], BOUNDS['temp'][0], BOUNDS['temp'][1]))
    z = int(np.clip(position[2], BOUNDS['moisture'][0], BOUNDS['moisture'][1]))
    return category1_multivariable_control(x, temp, z)


# === 目标函数 ===
def category1_multivariable_control(special_event, emergency_vehicles, visibility):
    triggered = set()

    if (special_event < 50) != (special_event < 70):
        triggered.add(1)
    if (special_event > 50) != (special_event > 30):
        triggered.add(2)
    if (special_event > 50) != (special_event > 70):
        triggered.add(3)
    if (special_event <= 25) != (special_event <= 75):
        triggered.add(4)
    if (special_event >= 25) != (special_event >= 75):
        triggered.add(5)
    if (special_event <= 50) != (special_event <= 25):
        triggered.add(6)

    if (emergency_vehicles < 200) != (emergency_vehicles < 100):
        triggered.add(7)
    if (emergency_vehicles < 200) != (emergency_vehicles < 300):
        triggered.add(8)
    if (emergency_vehicles > 200) != (emergency_vehicles > 100):
        triggered.add(9)
    if (emergency_vehicles > 200) != (emergency_vehicles > 300):
        triggered.add(10)
    if (emergency_vehicles <= 150) != (emergency_vehicles <= 250):
        triggered.add(11)
    if (emergency_vehicles >= 150) != (emergency_vehicles >= 250):
        triggered.add(12)
    if (emergency_vehicles % 2 == 0) != (emergency_vehicles % 3 == 0):
        triggered.add(13)

    if (visibility < 500) != (visibility < 300):
        triggered.add(14)
    if (visibility < 500) != (visibility < 700):
        triggered.add(15)
    if (visibility > 500) != (visibility > 300):
        triggered.add(16)
    if (visibility > 500) != (visibility > 700):
        triggered.add(17)
    if (visibility <= 400) != (visibility <= 600):
        triggered.add(18)
    if (visibility >= 400) != (visibility >= 600):
        triggered.add(19)
    if (visibility % 10 < 5) != (visibility % 20 < 5):
        triggered.add(20)

    if (special_event < 50 and emergency_vehicles < 200) != (
            special_event < 30 and emergency_vehicles < 200):
        triggered.add(21)
    if (special_event < 50 and emergency_vehicles < 200) != (
            special_event < 50 and emergency_vehicles < 150):
        triggered.add(22)
    if (special_event > 50 and emergency_vehicles > 200) != (
            special_event > 30 and emergency_vehicles > 200):
        triggered.add(23)
    if (special_event > 50 and emergency_vehicles > 200) != (
            special_event > 50 and emergency_vehicles > 150):
        triggered.add(24)
    if (special_event <= 40 or emergency_vehicles <= 180) != (
            special_event <= 60 or emergency_vehicles <= 180):
        triggered.add(25)
    if (special_event <= 40 or emergency_vehicles <= 180) != (
            special_event <= 40 or emergency_vehicles <= 320):
        triggered.add(26)

    if (special_event < 50 and visibility < 500) != (special_event < 30 and visibility < 500):
        triggered.add(27)
    if (special_event < 50 and visibility < 500) != (special_event < 50 and visibility < 700):
        triggered.add(28)
    if (special_event > 50 and visibility > 500) != (special_event > 30 and visibility > 500):
        triggered.add(29)
    if (special_event > 50 and visibility > 500) != (special_event > 50 and visibility > 200):
        triggered.add(30)
    if (special_event <= 60 or visibility <= 600) != (special_event <= 40 or visibility <= 600):
        triggered.add(31)
    if (special_event <= 60 or visibility <= 600) != (special_event <= 60 or visibility <= 400):
        triggered.add(32)

    if (emergency_vehicles < 200 and visibility < 500) != (
            emergency_vehicles < 150 and visibility < 500):
        triggered.add(33)
    if (emergency_vehicles < 200 and visibility < 500) != (
            emergency_vehicles < 200 and visibility < 800):
        triggered.add(34)
    if (emergency_vehicles > 200 and visibility > 500) != (
            emergency_vehicles > 150 and visibility > 500):
        triggered.add(35)
    if (emergency_vehicles > 200 and visibility > 500) != (
            emergency_vehicles > 200 and visibility > 200):
        triggered.add(36)
    if (emergency_vehicles <= 350 or visibility <= 700) != (
            emergency_vehicles <= 180 or visibility <= 700):
        triggered.add(37)
    if (emergency_vehicles <= 250 or visibility <= 700) != (
            emergency_vehicles <= 250 or visibility <= 500):
        triggered.add(38)

    if (special_event < 40 or emergency_vehicles > 300) != (
            special_event < 60 or emergency_vehicles > 300):
        triggered.add(39)
    if (special_event < 40 or emergency_vehicles > 300) != (
            special_event < 40 or emergency_vehicles > 250):
        triggered.add(40)
    if (special_event > 60 or visibility < 400) != (special_event > 40 or visibility < 400):
        triggered.add(41)
    if (special_event > 60 or visibility < 400) != (special_event > 60 or visibility < 300):
        triggered.add(42)
    if (emergency_vehicles < 150 or visibility > 700) != (emergency_vehicles < 200 or visibility > 700):
        triggered.add(43)
    if (emergency_vehicles < 150 or visibility > 700) != (emergency_vehicles < 150 or visibility > 400):
        triggered.add(44)

    if (special_event + emergency_vehicles < 350) != (special_event + emergency_vehicles < 200):
        triggered.add(45)
    if (special_event * 2 < emergency_vehicles) != (special_event * 3 < emergency_vehicles):
        triggered.add(46)
    if (visibility - special_event > 400) != (visibility - special_event > 500):
        triggered.add(47)
    if (emergency_vehicles % 50 == 0) != (emergency_vehicles % 100 == 0):
        triggered.add(48)
    if (special_event + visibility > 600) != (special_event + visibility > 500):
        triggered.add(49)
    if (abs(special_event - 50) < 20) != (abs(special_event - 50) < 30):
        triggered.add(50)

    if (special_event < 60 and emergency_vehicles < 250 and visibility > 400) != (
            special_event < 40 and emergency_vehicles < 250 and visibility > 400):
        triggered.add(51)
    if (special_event < 60 and emergency_vehicles < 250 and visibility > 400) != (
            special_event < 60 and emergency_vehicles < 200 and visibility > 400):
        triggered.add(52)
    if (special_event < 60 and emergency_vehicles < 250 and visibility > 400) != (
            special_event < 60 and emergency_vehicles < 250 and visibility > 300):
        triggered.add(53)

    if (special_event > 40 and emergency_vehicles > 150 and visibility < 600) != (
            special_event > 60 and emergency_vehicles > 150 and visibility < 600):
        triggered.add(54)
    if (special_event > 40 and emergency_vehicles > 150 and visibility < 600) != (
            special_event > 40 and emergency_vehicles > 200 and visibility < 600):
        triggered.add(55)
    if (special_event > 40 and emergency_vehicles > 150 and visibility < 600) != (
            special_event > 40 and emergency_vehicles > 150 and visibility < 500):
        triggered.add(56)

    if (special_event < 30 or emergency_vehicles < 100 or visibility > 800) != (
            special_event < 50 or emergency_vehicles < 100 or visibility > 800):
        triggered.add(57)
    if (special_event < 30 or emergency_vehicles < 100 or visibility > 800) != (
            special_event < 30 or emergency_vehicles < 150 or visibility > 800):
        triggered.add(58)
    if (special_event < 30 or emergency_vehicles < 100 or visibility > 800) != (
            special_event < 30 or emergency_vehicles < 100 or visibility > 700):
        triggered.add(59)

    if (special_event > 70 or emergency_vehicles > 300 or visibility < 200) != (
            special_event > 50 or emergency_vehicles > 300 or visibility < 200):
        triggered.add(60)
    if (special_event > 70 or emergency_vehicles > 300 or visibility < 200) != (
            special_event > 70 or emergency_vehicles > 250 or visibility < 200):
        triggered.add(61)
    if (special_event > 70 or emergency_vehicles > 300 or visibility < 200) != (
            special_event > 70 or emergency_vehicles > 300 or visibility < 300):
        triggered.add(62)

    if ((special_event < 50 and emergency_vehicles < 200) or visibility > 700) != (
            (special_event < 30 and emergency_vehicles < 200) or visibility > 700):
        triggered.add(63)
    if ((special_event < 50 and emergency_vehicles < 200) or visibility > 700) != (
            (special_event < 50 and emergency_vehicles < 150) or visibility > 700):
        triggered.add(64)
    if ((special_event < 50 and emergency_vehicles < 200) or visibility > 700) != (
            (special_event < 50 and emergency_vehicles < 200) or visibility > 400):
        triggered.add(65)

    if ((special_event > 50 or emergency_vehicles > 200) and visibility < 500) != (
            (special_event > 30 or emergency_vehicles > 200) and visibility < 500):
        triggered.add(66)
    if ((special_event > 50 or emergency_vehicles > 200) and visibility < 500) != (
            (special_event > 50 or emergency_vehicles > 150) and visibility < 500):
        triggered.add(67)
    if ((special_event > 50 or emergency_vehicles > 200) and visibility < 500) != (
            (special_event > 50 or emergency_vehicles > 200) and visibility < 400):
        triggered.add(68)

    if (special_event + emergency_vehicles > 300 and visibility < 600) != (
            special_event + emergency_vehicles > 250 and visibility < 600):
        triggered.add(69)
    if (special_event + emergency_vehicles > 300 and visibility < 600) != (
            special_event + emergency_vehicles > 300 and visibility < 400):
        triggered.add(70)

    if (special_event * 4 < visibility and emergency_vehicles > 100) != (
            special_event * 5 < visibility and emergency_vehicles > 100):
        triggered.add(71)
    if (special_event * 4 < visibility and emergency_vehicles > 100) != (
            special_event * 4 < visibility and emergency_vehicles > 150):
        triggered.add(72)

    if (visibility - emergency_vehicles > 200 and special_event > 30) != (
            visibility - emergency_vehicles > 300 and special_event > 30):
        triggered.add(73)
    if (visibility - emergency_vehicles > 200 and special_event > 30) != (
            visibility - emergency_vehicles > 200 and special_event > 50):
        triggered.add(74)

    if (special_event % 3 == 0 and emergency_vehicles > 120 and visibility <= 750) != (
            special_event % 4 == 0 and emergency_vehicles > 120 and visibility <= 750):
        triggered.add(75)
    if (special_event % 3 == 0 and emergency_vehicles > 120 and visibility <= 750) != (
            special_event % 3 == 0 and emergency_vehicles > 180 and visibility <= 750):
        triggered.add(76)
    if (special_event % 3 == 0 and emergency_vehicles > 120 and visibility <= 750) != (
            special_event % 3 == 0 and emergency_vehicles > 120 and visibility <= 650):
        triggered.add(77)

    if (abs(special_event - 50) > 10 and emergency_vehicles < 300 and visibility % 100 < 50) != (
            abs(special_event - 50) > 20 and emergency_vehicles < 300 and visibility % 100 < 50):
        triggered.add(78)
    if (abs(special_event - 50) > 10 and emergency_vehicles < 300 and visibility % 100 < 50) != (
            abs(special_event - 50) > 10 and emergency_vehicles < 250 and visibility % 100 < 50):
        triggered.add(79)
    if (abs(special_event - 50) > 10 and emergency_vehicles < 300 and visibility % 100 < 50) != (
            abs(special_event - 50) > 10 and emergency_vehicles < 300 and visibility % 100 < 30):
        triggered.add(80)

    if (20 <= special_event <= 80 and emergency_vehicles < 250) != (
            30 <= special_event <= 80 and emergency_vehicles < 250):
        triggered.add(81)
    if (20 <= special_event <= 80 and emergency_vehicles < 250) != (
            20 <= special_event <= 70 and emergency_vehicles < 250):
        triggered.add(82)
    if (20 <= special_event <= 80 and emergency_vehicles < 250) != (
            20 <= special_event <= 80 and emergency_vehicles < 200):
        triggered.add(83)

    if (100 <= emergency_vehicles <= 300 and visibility > 400) != (
            150 <= emergency_vehicles <= 300 and visibility > 400):
        triggered.add(84)
    if (100 <= emergency_vehicles <= 300 and visibility > 400) != (
            100 <= emergency_vehicles <= 250 and visibility > 400):
        triggered.add(85)
    if (100 <= emergency_vehicles <= 300 and visibility > 400) != (
            100 <= emergency_vehicles <= 300 and visibility > 300):
        triggered.add(86)

    if (300 <= visibility <= 700 and special_event > 25) != (
            400 <= visibility <= 700 and special_event > 25):
        triggered.add(87)
    if (300 <= visibility <= 700 and special_event > 25) != (
            300 <= visibility <= 600 and special_event > 25):
        triggered.add(88)
    if (300 <= visibility <= 700 and special_event > 25) != (
            300 <= visibility <= 700 and special_event > 35):
        triggered.add(89)

    if ((10 <= special_event <= 40) and (150 <= emergency_vehicles <= 350)) != (
            (20 <= special_event <= 40) and (150 <= emergency_vehicles <= 350)):
        triggered.add(90)
    if ((10 <= special_event <= 40) and (150 <= emergency_vehicles <= 350)) != (
            (10 <= special_event <= 30) and (150 <= emergency_vehicles <= 350)):
        triggered.add(91)
    if ((10 <= special_event <= 40) and (150 <= emergency_vehicles <= 350)) != (
            (10 <= special_event <= 40) and (200 <= emergency_vehicles <= 350)):
        triggered.add(92)
    if ((10 <= special_event <= 40) and (150 <= emergency_vehicles <= 350)) != (
            (10 <= special_event <= 40) and (150 <= emergency_vehicles <= 300)):
        triggered.add(93)

    if ((special_event > 60) and (200 < visibility < 800)) != (
            (special_event > 50) and (200 < visibility < 800)):
        triggered.add(94)
    if ((special_event > 60) and (200 < visibility < 800)) != (
            (special_event > 60) and (300 < visibility < 800)):
        triggered.add(95)
    if ((special_event > 60) and (200 < visibility < 800)) != (
            (special_event > 60) and (200 < visibility < 700)):
        triggered.add(96)

    if (special_event < 45 and (emergency_vehicles > 180 or visibility < 450)) != (
            special_event < 55 and (emergency_vehicles > 180 or visibility < 450)):
        triggered.add(97)
    if (special_event < 45 and (emergency_vehicles > 180 or visibility < 450)) != (
            special_event < 45 and (emergency_vehicles > 220 or visibility < 450)):
        triggered.add(98)
    if (special_event < 45 and (emergency_vehicles > 180 or visibility < 450)) != (
            special_event < 45 and (emergency_vehicles > 180 or visibility < 350)):
        triggered.add(99)

    if (emergency_vehicles > 175 and (special_event < 35 or visibility > 650)) != (
            emergency_vehicles > 225 and (special_event < 35 or visibility > 650)):
        triggered.add(100)
    if (emergency_vehicles > 175 and (special_event < 35 or visibility > 650)) != (
            emergency_vehicles > 175 and (special_event < 45 or visibility > 650)):
        triggered.add(101)
    if (emergency_vehicles > 175 and (special_event < 35 or visibility > 650)) != (
            emergency_vehicles > 175 and (special_event < 35 or visibility > 550)):
        triggered.add(102)

    if (special_event + emergency_vehicles // 2 > 150) != (
            special_event + emergency_vehicles // 3 > 150):
        triggered.add(103)
    if (visibility // 10 + special_event > 80) != (visibility // 15 + special_event > 80):
        triggered.add(104)
    if (emergency_vehicles % 10 + special_event % 10 < 15) != (
            emergency_vehicles % 15 + special_event % 10 < 15):
        triggered.add(105)
    if (emergency_vehicles % 10 + special_event % 10 < 15) != (
            emergency_vehicles % 10 + special_event % 15 < 15):
        triggered.add(106)

    if (max(special_event, emergency_vehicles // 4) > 45) != (
            max(special_event, emergency_vehicles // 5) > 45):
        triggered.add(107)
    if (min(visibility // 10, special_event) < 35) != (min(visibility // 15, special_event) < 35):
        triggered.add(108)
    if (min(visibility // 10, special_event) < 35) != (min(visibility // 10, special_event) < 25):
        triggered.add(109)
    if (special_event > visibility // 20) != (special_event > visibility // 25):
        triggered.add(110)

    if (special_event % 2 == emergency_vehicles % 2) != (special_event % 3 == emergency_vehicles % 2):
        triggered.add(111)
    if (special_event % 2 == emergency_vehicles % 2) != (special_event % 2 == emergency_vehicles % 3):
        triggered.add(112)
    if (visibility % 7 == 0) != (visibility % 11 == 0):
        triggered.add(113)
    if ((special_event + emergency_vehicles) % 5 == 0) != (
            (special_event + emergency_vehicles) % 7 == 0):
        triggered.add(114)

    if (special_event & 1 == visibility & 1) != (special_event & 3 == visibility & 1):
        triggered.add(115)
    if (special_event & 1 == visibility & 1) != (special_event & 1 == visibility & 3):
        triggered.add(116)
    if (emergency_vehicles >> 2 > special_event) != (emergency_vehicles >> 3 > special_event):
        triggered.add(117)
    if ((special_event ^ emergency_vehicles) % 50 < 25) != (
            (special_event ^ emergency_vehicles) % 60 < 25):
        triggered.add(118)
    if ((special_event ^ emergency_vehicles) % 50 < 25) != (
            (special_event ^ emergency_vehicles) % 50 < 30):
        triggered.add(119)

    if (len(str(special_event)) + len(str(emergency_vehicles)) > 4) != (
            len(str(special_event)) + len(str(emergency_vehicles)) > 5):
        triggered.add(120)
    if (sum(int(d) for d in str(visibility)) > 10) != (sum(int(d) for d in str(visibility)) > 12):
        triggered.add(121)
    if (special_event ** 2 % 100 < 50) != (special_event ** 2 % 120 < 50):
        triggered.add(122)
    if (special_event ** 2 % 100 < 50) != (special_event ** 2 % 100 < 60):
        triggered.add(123)

    if (abs(special_event - emergency_vehicles // 4) < 20) != (
            abs(special_event - emergency_vehicles // 5) < 20):
        triggered.add(124)
    if (abs(special_event - emergency_vehicles // 4) < 20) != (
            abs(special_event - emergency_vehicles // 4) < 25):
        triggered.add(125)
    if ((visibility + special_event + emergency_vehicles) % 100 > 50) != (
            (visibility + special_event + emergency_vehicles) % 120 > 50):
        triggered.add(126)

    return triggered


# 目标路径定义
targetPaths = [
    [2, 4, 5, 6, 8, 10, 11, 12, 15, 17, 18, 19, 20, 23, 28, 29, 44, 45, 49, 52, 57, 65, 74, 83, 89, 91, 98, 101, 102, 104, 105, 107, 111, 112, 113, 115, 117, 118, 120],
    [2, 4, 5, 6, 7, 9, 11, 12, 13, 14, 16, 18, 19, 21, 22, 27, 33, 43, 44, 45, 47, 57, 63, 64, 66, 67, 73, 74, 76, 91, 92, 104, 108, 112, 116, 117, 118, 120, 121],
    [2, 4, 5, 6, 7, 9, 11, 12, 13, 14, 16, 20, 21, 22, 27, 33, 42, 43, 45, 53, 57, 63, 64, 66, 67, 75, 86, 87, 89, 91, 92, 100, 109, 112, 114, 117, 118, 120, 122],
    [2, 4, 5, 6, 7, 9, 11, 12, 13, 14, 16, 18, 19, 21, 22, 27, 33, 43, 44, 47, 49, 57, 63, 64, 66, 67, 74, 76, 91, 92, 104, 108, 112, 114, 117, 118, 120],
    [7, 9, 11, 12, 13, 15, 17, 18, 19, 20, 22, 28, 34, 35, 43, 44, 45, 47, 49, 50, 64, 75, 81, 92, 98, 100, 110, 111, 112, 118, 120, 122, 124, 125, 126],
    [2, 4, 5, 6, 8, 10, 11, 12, 20, 23, 25, 26, 29, 31, 37, 39, 41, 45, 51, 52, 74, 83, 100, 103, 105, 106, 107, 110, 113, 114, 115, 117, 118, 120],
    [2, 4, 5, 6, 8, 9, 11, 12, 13, 15, 16, 18, 19, 28, 44, 45, 47, 49, 52, 57, 65, 73, 74, 78, 83, 91, 98, 101, 104, 107, 108, 112, 117, 120],
    [1, 3, 4, 5, 8, 10, 13, 14, 16, 18, 19, 26, 30, 32, 36, 40, 44, 45, 49, 60, 61, 65, 68, 70, 78, 79, 80, 85, 108, 111, 112, 116, 117, 120],
    [2, 4, 5, 6, 8, 10, 11, 12, 14, 16, 23, 25, 26, 27, 36, 39, 42, 45, 53, 54, 57, 69, 75, 83, 86, 87, 97, 103, 109, 111, 117, 118, 120, 122],
    [1, 3, 4, 5, 8, 10, 11, 12, 14, 16, 25, 26, 30, 36, 39, 42, 45, 53, 54, 60, 69, 75, 83, 86, 87, 94, 103, 104, 109, 111, 113, 117, 120, 121],
    [2, 4, 5, 6, 8, 10, 11, 12, 13, 20, 23, 25, 26, 29, 31, 37, 39, 41, 45, 51, 52, 57, 59, 74, 83, 97, 103, 111, 112, 113, 117, 118, 120, 122],
    [2, 4, 5, 6, 7, 9, 12, 15, 17, 18, 19, 20, 21, 22, 28, 29, 34, 43, 44, 48, 57, 63, 64, 72, 74, 76, 91, 92, 104, 116, 117, 119, 120],
    [1, 3, 4, 5, 8, 10, 11, 12, 14, 16, 25, 26, 30, 36, 39, 42, 45, 53, 54, 60, 69, 75, 83, 86, 87, 94, 97, 103, 104, 105, 106, 108, 120],
    [7, 9, 11, 12, 13, 15, 17, 22, 28, 34, 35, 43, 44, 45, 50, 64, 75, 77, 81, 92, 98, 100, 104, 106, 107, 111, 112, 116, 118, 120, 122, 124, 126],
    [8, 10, 11, 12, 14, 16, 18, 19, 26, 30, 32, 36, 44, 45, 49, 50, 65, 68, 70, 73, 75, 80, 82, 83, 108, 111, 113, 118, 120, 122, 124, 126],
    [4, 5, 6, 7, 9, 11, 12, 13, 14, 16, 18, 19, 22, 33, 43, 44, 50, 64, 67, 81, 89, 92, 99, 109, 112, 113, 115, 117, 119, 120],
    [1, 3, 4, 5, 8, 10, 11, 12, 13, 25, 26, 30, 36, 39, 45, 54, 60, 62, 69, 75, 83, 94, 103, 104, 105, 106, 109, 112, 118, 120],
    [1, 3, 4, 5, 13, 15, 17, 18, 19, 25, 38, 41, 44, 47, 49, 54, 56, 65, 70, 75, 94, 97, 108, 112, 113, 115, 116, 117, 120, 121],
    [7, 9, 11, 12, 14, 16, 18, 19, 24, 30, 32, 33, 43, 44, 45, 49, 55, 65, 68, 69, 71, 73, 75, 80, 103, 108, 111, 113, 118],
    [7, 9, 11, 12, 13, 15, 17, 22, 28, 34, 35, 43, 44, 64, 75, 90, 92, 104, 105, 111, 112, 119, 120, 122, 123, 124, 125, 126],
    [7, 9, 11, 12, 13, 20, 24, 26, 30, 33, 43, 45, 46, 55, 69, 75, 80, 95, 103, 106, 111, 112, 114, 120, 122],
    [7, 9, 11, 12, 24, 34, 35, 45, 46, 50, 59, 75, 76, 77, 82, 96, 103, 105, 106, 113, 115, 120, 121],
    [2, 4, 5, 6, 7, 9, 21, 29, 31, 39, 41, 51, 72, 74, 84, 106, 110, 111, 113, 114, 118, 120, 122],
    [7, 9, 13, 15, 17, 34, 45, 58, 65, 72, 80, 84, 88, 111, 112, 120, 126],
    [4, 5, 6, 13, 37, 48, 50, 93, 103, 109, 112, 114, 120, 122]
]


# === 增强版指标收集器 ===
class MetricsCollector:
    def __init__(self):
        self.pso_start_time = None
        self.pso_end_time = None
        self.perfect_solutions_count = 0
        self.total_paths_count = 0
        self.final_fitness_scores = []
        self.pso_convergence_iterations = []
        self.pso_reset_counts = []
        self.path_execution_times = []

    def start_pso_phase(self):
        self.pso_start_time = time.time()

    def end_pso_phase(self):
        self.pso_end_time = time.time()

    def record_pso_result(self, fitness, is_perfect_match, convergence_iter=None, path_id=None, method='MI-PSO',
                          reset_count=0, execution_time=0):
        self.final_fitness_scores.append(fitness)
        self.total_paths_count += 1
        self.pso_reset_counts.append(reset_count)
        self.path_execution_times.append(execution_time)

        if is_perfect_match:
            self.perfect_solutions_count += 1
        if convergence_iter is not None:
            self.pso_convergence_iterations.append(convergence_iter)


# === 粒子类 ===
class Particle:
    def __init__(self):
        self.position = np.array([
            np.random.uniform(BOUNDS['light'][0], BOUNDS['light'][1]),
            np.random.uniform(BOUNDS['temp'][0], BOUNDS['temp'][1]),
            np.random.uniform(BOUNDS['moisture'][0], BOUNDS['moisture'][1])
        ])
        self.velocity = np.array([
            np.random.uniform(-5, 5),
            np.random.uniform(-3, 3),
            np.random.uniform(-5, 5)
        ])
        self.best_position = self.position.copy()
        self.best_fitness = 0
        self.fitness = 0


# === 单独的 MI-PSO 优化器类（已完全修复随机性重评估Bug） ===
class MIPSO:
    def __init__(self, target_path, swarm_size=20, max_iterations=3000, g2=10, th_cv=1.2):
        self.target_path = target_path
        self.swarm_size = swarm_size
        self.max_iterations = max_iterations
        self.g2 = g2
        self.th_cv = th_cv

        # 完全随机初始化
        self.particles = [Particle() for _ in range(swarm_size)]
        self.global_best_position = None
        self.global_best_fitness = 0
        self.global_best_triggered = set()  # 核心修复：记录最高分对应的真实触发路径
        self.reset_count = 0

        # 初始化适应度评估
        for particle in self.particles:
            fit, trig = self.evaluate(particle.position)
            particle.fitness = fit
            particle.best_fitness = fit
            particle.best_position = particle.position.copy()

            if fit > self.global_best_fitness:
                self.global_best_fitness = fit
                self.global_best_position = particle.position.copy()
                self.global_best_triggered = trig

    def evaluate(self, position):
        """同时返回适应度和实际触发的路径"""
        try:
            triggered = execute_Tr(position)
            if self.target_path.issubset(triggered):
                return 1.0, triggered
            intersection = len(triggered & self.target_path)
            union = len(triggered | self.target_path)
            return (intersection / union if union > 0 else 0.0), triggered
        except:
            return 0.0, set()

    def calculate_cv(self, scores):
        """实现公式 25：综合变异系数"""
        N = self.swarm_size
        mean_fit = np.mean(scores)
        if mean_fit == 0: return float('inf')

        var_term = np.sum((scores - mean_fit) ** 2) / N
        adj_diff = sum(abs(scores[2 * i + 1] - scores[2 * i]) for i in range(N // 2))
        adj_term = ((2 / N) * adj_diff) ** 2

        numerator = math.sqrt(var_term + adj_term)
        denominator = 2 * np.sum(scores)

        if denominator == 0: return float('inf')
        return numerator / denominator

    def flip_mutation(self, position, rho=0.1):
        """实现公式 27：变异反转操作"""
        new_pos = np.copy(position)
        new_pos[0] = BOUNDS['light'][0] + BOUNDS['light'][1] - position[0] + rho
        new_pos[1] = BOUNDS['temp'][0] + BOUNDS['temp'][1] - position[1] + rho
        new_pos[2] = BOUNDS['moisture'][0] + BOUNDS['moisture'][1] - position[2] + rho
        return clip_state(new_pos)

    def optimize(self):
        w = 0.7
        c1 = 1.5
        c2 = 1.5

        max_velocity = np.array([
            (BOUNDS['light'][1] - BOUNDS['light'][0]) * 0.2,
            (BOUNDS['temp'][1] - BOUNDS['temp'][0]) * 0.2,
            (BOUNDS['moisture'][1] - BOUNDS['moisture'][0]) * 0.2
        ])

        c = 0
        early_stop = False
        converged_at = self.max_iterations

        for iteration in range(self.max_iterations):
            scores = np.array([p.fitness for p in self.particles])
            c += 1

            # MI-PSO 特有的停滞检测与变异 (对应 Algorithm 4 Line 20-23)
            if c == self.g2:
                cv = self.calculate_cv(scores)
                if cv <= self.th_cv:
                    self.reset_count += 1
                    l = int(round(self.swarm_size * (1 - cv)))
                    l = min(max(l, 1), self.swarm_size)

                    selected_indices = np.random.choice(self.swarm_size, l, replace=False)
                    for idx in selected_indices:
                        new_pos = self.flip_mutation(self.particles[idx].position)
                        new_fit, new_trig = self.evaluate(new_pos)

                        if new_fit > self.particles[idx].fitness:
                            self.particles[idx].position = new_pos
                            self.particles[idx].fitness = new_fit
                            if new_fit > self.particles[idx].best_fitness:
                                self.particles[idx].best_fitness = new_fit
                                self.particles[idx].best_position = new_pos.copy()
                            if new_fit > self.global_best_fitness:
                                self.global_best_fitness = new_fit
                                self.global_best_position = new_pos.copy()
                                self.global_best_triggered = new_trig  # 更新最优路径
                c = 0

                # 标准 PSO 速度与位置更新
            for particle in self.particles:
                r1, r2 = np.random.random(3), np.random.random(3)
                particle.velocity = (w * particle.velocity +
                                     c1 * r1 * (particle.best_position - particle.position) +
                                     c2 * r2 * (self.global_best_position - particle.position))
                particle.velocity = np.clip(particle.velocity, -max_velocity, max_velocity)

                particle.position += particle.velocity
                particle.position = clip_state(particle.position)

                fit, trig = self.evaluate(particle.position)
                particle.fitness = fit

                if fit > particle.best_fitness:
                    particle.best_fitness = fit
                    particle.best_position = particle.position.copy()
                if fit > self.global_best_fitness:
                    self.global_best_fitness = fit
                    self.global_best_position = particle.position.copy()
                    self.global_best_triggered = trig  # 更新最优路径

            if self.global_best_fitness >= 1.0:
                converged_at = iteration + 1
                early_stop = True
                break

        return self.global_best_position, self.global_best_fitness, self.global_best_triggered, converged_at, early_stop, self.reset_count


# === Excel导出函数 ===
def export_multiple_runs_to_excel(all_run_results, all_run_metrics, num_runs, filename=None):
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Isolated_MIPSO_{num_runs}Runs_{timestamp}.xlsx"

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    border = Border(
        left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 工作表1: 运行汇总
    ws1 = wb.active
    ws1.title = "运行汇总"
    ws1.sheet_view.showGridLines = False

    headers = ["运行次数", "成功率", "成功数量", "平均适应度", "平均迭代次数", "运行时间(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, border, center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, (results, run_metrics) in enumerate(zip(all_run_results, all_run_metrics), start=1):
        success_count = sum(1 for r in results if r['perfect_match'])
        success_rate = (success_count / len(targetPaths)) * 100
        avg_fitness = np.mean([r['fitness'] for r in results])

        iterations_list = [r['convergence_iteration'] if r.get('convergence_iteration') is not None else 10000 for r in
                           results]
        avg_iterations = np.mean(iterations_list)
        total_time = run_metrics.pso_end_time - run_metrics.pso_start_time if run_metrics.pso_end_time else 0

        row_data = [f"运行 {run_idx}", f"{success_rate:.1f}%", f"{success_count}/{len(targetPaths)}",
                    f"{avg_fitness:.4f}", f"{avg_iterations:.1f}", f"{total_time:.2f}"]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border, cell.alignment = border, center_align
            if run_idx % 2 == 0: cell.fill = alternate_fill
            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    # 工作表2: 路径统计
    ws2 = wb.create_sheet(title="路径统计")
    ws2.sheet_view.showGridLines = False
    headers2 = ["路径编号", "成功次数", "成功率", "平均适应度", "平均迭代次数", "最小迭代", "最大迭代"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, border, center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(targetPaths)):
        success_count = sum(1 for results in all_run_results if results[path_idx]['perfect_match'])
        success_rate = (success_count / num_runs) * 100
        avg_fitness = np.mean([results[path_idx]['fitness'] for results in all_run_results])

        iterations_list = [results[path_idx]['convergence_iteration'] if results[path_idx].get(
            'convergence_iteration') is not None else 10000 for results in all_run_results]
        avg_iterations = np.mean(iterations_list)

        row_data = [f"路径 {path_idx + 1}", f"{success_count}/{num_runs}", f"{success_rate:.1f}%",
                    f"{avg_fitness:.4f}", f"{avg_iterations:.1f}", np.min(iterations_list), np.max(iterations_list)]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border, cell.alignment = border, center_align
            if (path_idx + 1) % 2 == 0: cell.fill = alternate_fill
            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    # 工作表3: 最佳粒子详情
    ws3 = wb.create_sheet(title="最佳粒子详情")
    ws3.sheet_view.showGridLines = False
    headers3 = ["路径", "运行", "最佳粒子(light,temp,moisture)", "适应度", "迭代次数", "求解方法", "生成路径"]
    col_widths3 = [10, 10, 25, 12, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, border, center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(targetPaths)):
        for run_idx, results in enumerate(all_run_results, start=1):
            result = results[path_idx]
            particle_str = f"({int(result['best_position'][0])}, {int(result['best_position'][1])}, {int(result['best_position'][2])})"
            path_str = str(sorted(list(result['triggered'])))

            convergence_iter = result['convergence_iteration'] if result.get(
                'convergence_iteration') is not None else 10000

            row_data = [f"路径{path_idx + 1}", f"运行{run_idx}", particle_str, f"{result['fitness']:.4f}",
                        convergence_iter if convergence_iter < 10000 else "-", "MI-PSO", path_str]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = left_align if col == 7 else center_align

                if result['fitness'] == 1.0:
                    cell.fill = success_fill
                elif result['fitness'] < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill
            row_idx += 1

    # 保存
    output_dir = os.path.join(os.getcwd(), "results")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"\n✓ 独立 MI-PSO 消融实验结果已导出到: {filepath}")
    return filepath


def run_single_mipso_experiment(run_num):
    print(f"\n{'=' * 50}")
    print(f"开始第 {run_num} 次运行 (独立 MI-PSO)")
    print(f"{'=' * 50}")

    run_metrics = MetricsCollector()
    run_metrics.start_pso_phase()

    pso_results = []

    for i, target_path in enumerate(targetPaths):
        path_start_time = time.time()

        mipso = MIPSO(target_path, swarm_size=20, max_iterations=3000)
        # 直接接收并使用最佳寻优过程中真实触发的 best_trig，坚决不重新 evaluate
        best_pos, best_fit, best_trig, converged_at, early_stop, reset_count = mipso.optimize()

        path_execution_time = time.time() - path_start_time

        # 使用真实的最高分数判断是否完美
        is_perfect = (best_fit >= 1.0)

        pso_results.append({
            'target_path': target_path,
            'best_position': best_pos,
            'fitness': best_fit,
            'triggered': best_trig,
            'perfect_match': is_perfect,
            'method': 'MI-PSO',
            'convergence_iteration': converged_at,
            'early_stopped': early_stop,
            'reset_count': reset_count
        })

        run_metrics.record_pso_result(
            fitness=best_fit,
            is_perfect_match=is_perfect,
            convergence_iter=converged_at if early_stop else None,
            path_id=i + 1,
            method='MI-PSO',
            reset_count=reset_count,
            execution_time=path_execution_time
        )

        status = "✓完美(MI-PSO)" if is_perfect else f"○部分({best_fit:.3f})"
        print(f"  路径{i + 1}: {status} | 变异触发次数: {reset_count} | 耗时 {path_execution_time:.2f}s")

    run_metrics.end_pso_phase()

    success_count = sum(1 for r in pso_results if r['perfect_match'])
    success_rate = (success_count / len(targetPaths)) * 100
    pso_time = run_metrics.pso_end_time - run_metrics.pso_start_time

    print(f"\n第{run_num}次运行完成: 成功率 {success_rate:.1f}% | 耗时 {pso_time:.2f}秒")
    return pso_results, run_metrics


def run_multiple_mipso_experiments(num_runs):
    all_run_results = []
    all_run_metrics = []
    total_start = time.time()

    for run_num in range(1, num_runs + 1):
        results, metrics = run_single_mipso_experiment(run_num)
        all_run_results.append(results)
        all_run_metrics.append(metrics)

    total_time = time.time() - total_start
    print(f"\n{'=' * 50}\n全部{num_runs}次独立 MI-PSO 运行完成! 总耗时: {total_time:.2f}秒\n{'=' * 50}")

    return all_run_results, all_run_metrics


if __name__ == "__main__":
    if len(sys.argv) > 1:
        try:
            NUM_RUNS = int(sys.argv[1])
        except ValueError:
            pass

    print("=" * 70)
    print(" 独立 MI-PSO 消融实验 (无 DQN 介入，完美修复随机性 Bug)")
    print("=" * 70)

    all_run_results, all_run_metrics = run_multiple_mipso_experiments(NUM_RUNS)
    export_multiple_runs_to_excel(all_run_results, all_run_metrics, NUM_RUNS)