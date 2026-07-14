import os
import random
from collections import deque
import numpy as np
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# --- 全局状态范围配置 ---
STATE_MIN_X, STATE_MAX_X = 1, 6
STATE_MIN_Y, STATE_MAX_Y = 1, 6
STATE_MIN_Z, STATE_MAX_Z = 1, 60

def generate_input():
    return [
        random.randint(STATE_MIN_X, STATE_MAX_X),
        random.randint(STATE_MIN_Y, STATE_MAX_Y),
        random.randint(STATE_MIN_Z, STATE_MAX_Z)
    ]


def execute_Tr(weather, time_period, z):
    """执行验证规则并返回触发的分支"""
    triggered = set()

    # Fixed all if statements - using triggered.add() instead of b[0]=1
    if (weather == 1) != (weather == 2):
        triggered.add(1)
    if (weather == 2) != (weather == 3):
        triggered.add(2)
    if (weather == 3) != (weather == 4):
        triggered.add(3)
    if (weather == 4) != (weather == 5):
        triggered.add(4)
    if (weather == 5) != (weather == 6):
        triggered.add(5)
    if (weather == 6) != (weather == 1):
        triggered.add(6)

    # 时间段相关规则 (7-12)
    if (time_period == 1) != (time_period == 2):
        triggered.add(7)
    if (time_period == 2) != (time_period == 3):
        triggered.add(8)
    if (time_period == 3) != (time_period == 4):
        triggered.add(9)
    if (time_period == 4) != (time_period == 5):
        triggered.add(10)
    if (time_period == 5) != (time_period == 6):
        triggered.add(11)
    if (time_period == 6) != (time_period == 1):
        triggered.add(12)

    # 行人数量相关规则 (13-22)
    if (z < 20) != (z < 30):
        triggered.add(13)
    if (z < 30) != (z < 40):
        triggered.add(14)
    if (z < 40) != (z < 50):
        triggered.add(15)
    if (z > 20) != (z > 30):
        triggered.add(16)
    if (z > 30) != (z > 40):
        triggered.add(17)
    if (z > 40) != (z > 50):
        triggered.add(18)
    if (z > 50) != (z > 60):
        triggered.add(19)
    if (10 < z < 50) != (15 < z < 50):
        triggered.add(20)
    if (15 < z < 45) != (20 < z < 45):
        triggered.add(21)
    if (20 < z < 40) != (25 < z < 40):
        triggered.add(22)

    # 天气组合规则 (23-31)
    if (weather in [1, 2]) != (weather in [1, 3]):
        triggered.add(23)
    if (weather in [2, 3]) != (weather in [2, 4]):
        triggered.add(24)
    if (weather in [3, 4]) != (weather in [3, 5]):
        triggered.add(25)
    if (weather in [4, 5]) != (weather in [4, 6]):
        triggered.add(26)
    if (weather in [5, 6]) != (weather in [5, 1]):
        triggered.add(27)
    if (weather in [1, 3, 5]) != (weather in [2, 3, 5]):
        triggered.add(28)
    if (weather in [2, 4, 6]) != (weather in [2, 4, 1]):
        triggered.add(29)
    if (weather in [1, 2, 3]) != (weather in [1, 2, 4]):
        triggered.add(30)
    if (weather in [4, 5, 6]) != (weather in [3, 5, 6]):
        triggered.add(31)

    # 时间段组合规则 (32-38)
    if (time_period in [1, 2]) != (time_period in [1, 3]):
        triggered.add(32)
    if (time_period in [3, 4]) != (time_period in [3, 5]):
        triggered.add(33)
    if (time_period in [5, 6]) != (time_period in [5, 1]):
        triggered.add(34)
    if (time_period in [1, 3, 5]) != (time_period in [2, 3, 5]):
        triggered.add(35)
    if (time_period in [2, 4, 6]) != (time_period in [2, 4, 1]):
        triggered.add(36)
    if (time_period in [1, 2, 3]) != (time_period in [1, 2, 4]):
        triggered.add(37)
    if (time_period in [4, 5, 6]) != (time_period in [4, 5, 1]):
        triggered.add(38)

    # 天气和时间段交互规则 (39-50)
    if (weather == 1 and time_period in [1, 2]) != (weather == 2 and time_period in [1, 2]):
        triggered.add(39)
    if (weather == 1 and time_period in [3, 4]) != (weather == 1 and time_period in [3, 5]):
        triggered.add(40)
    if (weather == 1 and time_period in [5, 6]) != (weather == 1 and time_period in [5, 1]):
        triggered.add(41)
    if (weather == 2 and time_period in [1, 2]) != (weather == 3 and time_period in [1, 2]):
        triggered.add(42)
    if (weather == 2 and time_period in [3, 4]) != (weather == 2 and time_period in [3, 5]):
        triggered.add(43)
    if (weather == 2 and time_period in [5, 6]) != (weather == 2 and time_period in [5, 1]):
        triggered.add(44)
    if (weather == 3 and time_period in [1, 2]) != (weather == 4 and time_period in [1, 2]):
        triggered.add(45)
    if (weather == 3 and time_period in [3, 4]) != (weather == 3 and time_period in [3, 5]):
        triggered.add(46)
    if (weather == 3 and time_period in [5, 6]) != (weather == 3 and time_period in [5, 1]):
        triggered.add(47)
    if (weather == 4 and time_period in [1, 2]) != (weather == 5 and time_period in [1, 2]):
        triggered.add(48)
    if (weather == 4 and time_period in [3, 4]) != (weather == 4 and time_period in [3, 5]):
        triggered.add(49)
    if (weather == 4 and time_period in [5, 6]) != (weather == 4 and time_period in [5, 1]):
        triggered.add(50)

    # 天气和行人数量交互规则 (51-62)
    if (weather == 1 and z > 30) != (weather == 1 and z > 35):
        triggered.add(51)
    if (weather == 1 and z < 40) != (weather == 1 and z < 45):
        triggered.add(52)
    if (weather == 2 and z > 25) != (weather == 2 and z > 30):
        triggered.add(53)
    if (weather == 2 and z < 45) != (weather == 2 and z < 50):
        triggered.add(54)
    if (weather == 3 and z > 20) != (weather == 3 and z > 25):
        triggered.add(55)
    if (weather == 3 and z < 50) != (weather == 3 and z < 55):
        triggered.add(56)
    if (weather == 4 and z > 15) != (weather == 4 and z > 20):
        triggered.add(57)
    if (weather == 4 and z < 45) != (weather == 4 and z < 50):
        triggered.add(58)
    if (weather == 5 and z > 25) != (weather == 5 and z > 30):
        triggered.add(59)
    if (weather == 5 and z < 40) != (weather == 5 and z < 45):
        triggered.add(60)
    if (weather == 6 and z > 15) != (weather == 6 and z > 20):
        triggered.add(61)
    if (weather == 6 and z < 35) != (weather == 6 and z < 40):
        triggered.add(62)

    # 时间段和行人数量交互规则 (63-74)
    if (time_period == 1 and z > 35) != (time_period == 1 and z > 40):
        triggered.add(63)
    if (time_period == 1 and z < 45) != (time_period == 1 and z < 50):
        triggered.add(64)
    if (time_period == 2 and z > 30) != (time_period == 2 and z > 35):
        triggered.add(65)
    if (time_period == 2 and z < 50) != (time_period == 2 and z < 55):
        triggered.add(66)
    if (time_period == 3 and z > 40) != (time_period == 3 and z > 45):
        triggered.add(67)
    if (time_period == 3 and z < 35) != (time_period == 3 and z < 30):
        triggered.add(68)
    if (time_period == 4 and z > 20) != (time_period == 4 and z > 25):
        triggered.add(69)
    if (time_period == 4 and z < 30) != (time_period == 4 and z < 25):
        triggered.add(70)
    if (time_period == 5 and z > 45) != (time_period == 5 and z > 50):
        triggered.add(71)
    if (time_period == 5 and z < 25) != (time_period == 5 and z < 20):
        triggered.add(72)
    if (time_period == 6 and z > 50) != (time_period == 6 and z > 55):
        triggered.add(73)
    if (time_period == 6 and z < 20) != (time_period == 6 and z < 15):
        triggered.add(74)

    # 三元素组合规则 (75-84)
    if (weather in [1, 2] and time_period in [1, 2] and z > 30) != (
            weather in [1, 3] and time_period in [1, 2] and z > 30):
        triggered.add(75)
    if (weather in [1, 2] and time_period in [1, 2] and z < 40) != (
            weather in [1, 2] and time_period in [1, 3] and z < 40):
        triggered.add(76)
    if (weather in [3, 4] and time_period in [1, 2] and z > 25) != (
            weather in [3, 5] and time_period in [1, 2] and z > 25):
        triggered.add(77)
    if (weather in [3, 4] and time_period in [1, 2] and z < 35) != (
            weather in [3, 4] and time_period in [1, 3] and z < 35):
        triggered.add(78)
    if (weather in [5, 6] and time_period in [1, 2] and z > 20) != (
            weather in [5, 1] and time_period in [1, 2] and z > 20):
        triggered.add(79)
    if (weather in [5, 6] and time_period in [1, 2] and z < 30) != (
            weather in [5, 6] and time_period in [1, 3] and z < 30):
        triggered.add(80)
    if (weather in [1, 3] and time_period in [3, 4] and z > 35) != (
            weather in [1, 4] and time_period in [3, 4] and z > 35):
        triggered.add(81)
    if (weather in [2, 4] and time_period in [3, 4] and z > 30) != (
            weather in [2, 5] and time_period in [3, 4] and z > 30):
        triggered.add(82)
    if (weather in [1, 5] and time_period in [5, 6] and z > 40) != (
            weather in [1, 6] and time_period in [5, 6] and z > 40):
        triggered.add(83)
    if (weather in [2, 6] and time_period in [5, 6] and z > 25) != (
            weather in [2, 1] and time_period in [5, 6] and z > 25):
        triggered.add(84)

    # 复杂条件规则 (85-100)
    if (weather <= 3 and time_period <= 3 and z > 25) != (weather <= 4 and time_period <= 3 and z > 25):
        triggered.add(85)
    if (weather >= 4 and time_period >= 4 and z > 20) != (weather >= 3 and time_period >= 4 and z > 20):
        triggered.add(86)
    if (weather <= 2 and time_period >= 4 and z < 35) != (weather <= 3 and time_period >= 4 and z < 35):
        triggered.add(87)
    if (weather >= 5 and time_period <= 2 and z < 40) != (weather >= 4 and time_period <= 2 and z < 40):
        triggered.add(88)
    if (weather % 2 == 1 and time_period % 2 == 1) != (weather % 2 == 0 and time_period % 2 == 1):
        triggered.add(89)
    if (weather % 2 == 0 and time_period % 2 == 0) != (weather % 2 == 1 and time_period % 2 == 0):
        triggered.add(90)
    if (weather + time_period > 6) != (weather + time_period > 7):
        triggered.add(91)
    if (weather + time_period < 5) != (weather + time_period < 4):
        triggered.add(92)
    if (weather * time_period > 10) != (weather * time_period > 12):
        triggered.add(93)
    if (weather * time_period < 8) != (weather * time_period < 6):
        triggered.add(94)
    if (abs(weather - time_period) <= 2) != (abs(weather - time_period) <= 3):
        triggered.add(95)
    if (abs(weather - time_period) >= 3) != (abs(weather - time_period) >= 2):
        triggered.add(96)
    if (z % 10 < 5) != (z % 10 < 6):
        triggered.add(97)
    if (z % 10 >= 5) != (z % 10 >= 4):
        triggered.add(98)
    if (z // 10 >= 3) != (z // 10 >= 2):
        triggered.add(99)
    if (z // 10 <= 2) != (z // 10 <= 3):
        triggered.add(100)

    # 高级组合规则 (101-113)
    if ((weather + time_period + z // 10) % 3 == 0) != ((weather + time_period + z // 10) % 3 == 1):
        triggered.add(101)
    if ((weather + time_period + z // 10) % 3 == 1) != ((weather + time_period + z // 10) % 3 == 2):
        triggered.add(102)
    if ((weather + time_period + z // 10) % 3 == 2) != ((weather + time_period + z // 10) % 3 == 0):
        triggered.add(103)
    if (weather * time_period + z // 10 > 15) != (weather * time_period + z // 10 > 16):
        triggered.add(104)
    if (weather * time_period + z // 10 < 12) != (weather * time_period + z // 10 < 11):
        triggered.add(105)
    if ((weather * time_period) % (z // 10 + 1) == 0) != ((weather * time_period) % (z // 10 + 2) == 0):
        triggered.add(106)
    if (weather > time_period and z > 30) != (weather > time_period and z > 35):
        triggered.add(107)
    if (weather < time_period and z < 30) != (weather < time_period and z < 25):
        triggered.add(108)
    if (weather == time_period) != (weather == time_period + 1):
        triggered.add(109)
    if (weather + time_period == z // 10) != (weather + time_period == z // 10 + 1):
        triggered.add(110)
    if (abs(weather - time_period) == z // 10) != (abs(weather - time_period) == z // 10 + 1):
        triggered.add(111)
    if (max(weather, time_period) == z // 10) != (max(weather, time_period) == z // 10 + 1):
        triggered.add(112)
    if (min(weather, time_period) * 10 <= z) != (min(weather, time_period) * 11 <= z):
        triggered.add(113)

    return triggered


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0


# Fixed: Changed to proper list syntax (removed A1 =, etc.)
TARGET_PATHS = [
    [2, 3, 7, 12, 14, 17, 23, 24, 30, 31, 34, 35, 36, 38, 42, 45, 47, 75, 89, 92, 96, 97, 100, 101, 102, 107, 110, 112],
    [3, 4, 8, 9, 14, 17, 24, 25, 30, 31, 32, 37, 68, 78, 82, 85, 89, 91, 93, 100, 101, 102, 106, 107, 109, 112],
    [2, 3, 7, 12, 15, 18, 23, 24, 30, 31, 34, 35, 36, 38, 42, 45, 47, 64, 75, 89, 92, 96, 97, 102, 103, 110],
    [2, 3, 9, 10, 15, 18, 23, 24, 30, 31, 33, 37, 46, 81, 86, 90, 91, 93, 98, 102, 103, 104, 106, 112],
    [2, 3, 7, 12, 20, 23, 24, 30, 31, 34, 35, 36, 38, 42, 45, 47, 89, 92, 96, 98, 102, 103, 106, 111],
    [2, 3, 11, 12, 14, 16, 23, 24, 30, 31, 34, 36, 38, 47, 86, 87, 90, 95, 100, 101, 103, 111, 113],
    [5, 6, 7, 12, 19, 26, 27, 29, 34, 35, 36, 38, 79, 89, 91, 94, 97, 101, 103, 105, 106, 111, 112],
    [1, 6, 11, 12, 15, 18, 27, 28, 29, 34, 36, 38, 41, 52, 84, 90, 91, 94, 98, 102, 103, 106, 111],
    [3, 4, 7, 12, 20, 24, 25, 30, 31, 34, 35, 36, 38, 45, 48, 50, 88, 89, 95, 97, 101, 103, 106],
    [5, 6, 7, 12, 13, 16, 22, 26, 27, 29, 34, 35, 36, 38, 79, 89, 91, 94, 98, 99, 101, 103, 106],
    [4, 5, 7, 12, 14, 17, 25, 26, 34, 35, 36, 38, 48, 77, 89, 97, 100, 101, 103, 106, 107, 111],
    [2, 3, 7, 12, 18, 23, 24, 30, 31, 34, 35, 36, 38, 42, 45, 47, 56, 75, 89, 92, 96, 101, 103],
    [2, 3, 11, 12, 13, 21, 23, 24, 30, 31, 34, 36, 38, 47, 87, 90, 95, 99, 102, 103, 106, 111],
    [1, 6, 9, 10, 13, 16, 22, 27, 28, 29, 33, 37, 40, 69, 90, 95, 98, 99, 101, 102, 106, 111],
    [1, 6, 7, 12, 15, 18, 27, 28, 29, 34, 35, 36, 38, 39, 41, 64, 79, 89, 97, 101, 103, 109],
    [1, 2, 11, 12, 15, 18, 23, 28, 34, 36, 38, 44, 54, 90, 93, 97, 101, 103, 104, 106, 111],
    [3, 4, 11, 12, 13, 21, 24, 25, 30, 31, 34, 36, 38, 50, 57, 90, 96, 99, 101, 103, 111],
    [1, 6, 8, 9, 14, 16, 27, 28, 29, 32, 37, 68, 76, 89, 92, 96, 100, 101, 102, 110, 112],
    [5, 6, 8, 9, 13, 16, 22, 26, 27, 29, 32, 37, 80, 89, 95, 98, 99, 102, 103, 106, 111],
    [1, 6, 10, 11, 14, 17, 27, 28, 29, 33, 40, 51, 84, 89, 98, 100, 101, 103, 106, 111],
    [3, 4, 10, 11, 13, 16, 22, 24, 25, 30, 31, 33, 49, 89, 97, 99, 102, 103, 106, 108],
    [1, 2, 10, 11, 13, 16, 22, 23, 28, 33, 43, 72, 89, 91, 95, 99, 101, 103, 111, 113],
    [3, 4, 11, 12, 20, 24, 25, 30, 31, 34, 36, 38, 50, 74, 90, 96, 97, 102, 103, 111],
    [4, 5, 8, 9, 15, 18, 25, 26, 32, 37, 60, 67, 82, 89, 96, 98, 101, 103, 106, 112],
    [2, 3, 10, 11, 18, 23, 24, 30, 31, 33, 46, 56, 71, 86, 89, 96, 101, 102, 112],
    [4, 5, 10, 11, 15, 18, 25, 26, 33, 83, 89, 97, 102, 103, 106, 109, 112],
    [4, 5, 11, 12, 14, 16, 25, 26, 34, 36, 38, 59, 90, 100, 102, 103, 106],
    [3, 4, 7, 8, 14, 17, 24, 25, 30, 31, 32, 35, 45, 48, 65, 77, 85, 88, 90, 96, 97, 100, 101, 103, 105, 106, 107, 112],
    [3, 4, 7, 12, 15, 17, 24, 25, 30, 31, 34, 35, 36, 38, 45, 48, 50, 63, 77, 85, 89, 95, 101, 103, 110, 112],
    [2, 3, 8, 9, 13, 16, 22, 23, 24, 30, 31, 32, 37, 55, 78, 89, 97, 99, 102, 103, 105, 106, 109, 112],
    [1, 2, 9, 10, 13, 16, 22, 23, 28, 33, 37, 43, 69, 70, 90, 96, 97, 99, 102, 103, 106, 108, 111],
    [5, 6, 7, 12, 14, 17, 26, 27, 29, 34, 35, 36, 38, 62, 79, 89, 91, 94, 97, 100, 101, 102, 107],
    [1, 2, 9, 10, 14, 16, 23, 28, 33, 37, 43, 53, 90, 96, 100, 101, 103, 105, 106, 112],
    [3, 4, 7, 8, 18, 24, 25, 30, 31, 32, 35, 45, 48, 66, 77, 85, 90, 96, 102, 103, 110],
    [4, 5, 11, 12, 19, 25, 26, 34, 36, 38, 73, 83, 90, 98, 101, 102, 106, 112, 113],
    [3, 4, 10, 11, 15, 18, 24, 25, 30, 31, 33, 49, 58, 89, 97, 101, 102, 106, 112],
    [5, 6, 11, 12, 13, 21, 26, 27, 29, 34, 36, 38, 61, 90, 99, 102, 103, 109]
]


class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 2000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(TARGET_PATHS)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }


def compute_robustness(state, path):
    base = execute_Tr(*state)
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                    np.clip(state[0] + dx, STATE_MIN_X, STATE_MAX_X),
                    np.clip(state[1] + dy, STATE_MIN_Y, STATE_MAX_Y),
                    np.clip(state[2] + dz, STATE_MIN_Z, STATE_MAX_Z)
                ])
                n_trig = execute_Tr(*neighbor)
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = TARGET_PATHS[target_path_idx]
    samples = []
    attempts = 0

    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(STATE_MIN_X, STATE_MAX_X),
            random.randint(STATE_MIN_Y, STATE_MAX_Y),
            random.randint(STATE_MIN_Z, STATE_MAX_Z)
        ])
        triggered = execute_Tr(*state)

        if not triggered:
            continue

        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)

        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }

        samples.append(sample_data)

    return samples


def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample[
                'robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)

        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]

    return selected_samples


def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = TARGET_PATHS[target_path_idx]
        samples = []
        attempts = 0

        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(STATE_MIN_X, STATE_MAX_X),
                random.randint(STATE_MIN_Y, STATE_MAX_Y),
                random.randint(STATE_MIN_Z, STATE_MAX_Z)
            ])
            triggered = execute_Tr(*state)

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)

            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)

        random.shuffle(samples)
        return samples[:config.top_k_samples]

    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")

        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)


def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}

    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates

    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}

        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config,
                                                         shared_candidates[path_idx])

            strategy_results[path_idx] = samples

        results[strategy_name] = strategy_results

    return results


def analyze_fitness_values(results, config):
    analysis_results = {}

    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []

        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])

        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }

        all_scores = []

        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])

        scores_array = np.array(all_scores)

        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)

        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })

        analysis_results[strategy_name] = analysis

    return analysis_results


def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []

    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)

    df = pd.DataFrame(df_data)
    return df, analysis_results


def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []

    print(f"Starting {num_runs} experiments...")

    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")

        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)

        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)

        print(f"Experiment {run_idx + 1} completed.")

    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df


def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')

    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width

    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []

    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]

        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)

    stats_df = pd.DataFrame(stats_data)

    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Results saved to: {output_path}")


def main():
    results_df = run_multiple_experiments(num_runs=20)

    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")

    save_results_to_excel(results_df, output_path)

    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)

    return results_df, output_path


if __name__ == "__main__":
    results, output_file = main()