
import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
import numpy as np
import random
import time
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# ===  ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'MIN_X': 1, 'MAX_X': 100,
    'MIN_Y': 1, 'MAX_Y': 100,
    'MIN_Z': 1, 'MAX_Z': 100,
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 5,
    'NUM_ROUNDS': 5,
    'REPLAY_BATCH_SIZE': 64,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'LEARNING_RATE': 3e-4,
    'NUM_RUNS': 20,  # 20 run
    'TOP_K_SAMPLES': 20,
    'REPLAY_BUFFER_CAPACITY': 20000,  # Path 
    'TARGET_PATHS': [
        [2, 3, 7, 12, 14, 17, 23, 24, 30, 31, 34, 35, 36, 38, 42, 45, 47, 75, 89, 92, 96, 97, 100, 101, 102, 107, 110,
         112],
        [3, 4, 8, 9, 14, 17, 24, 25, 30, 31, 32, 37, 68, 78, 82, 85, 89, 91, 93, 100, 101, 102, 106, 107, 109, 112],
        [2, 3, 7, 12, 15, 18, 23, 24, 30, 31, 34, 35, 36, 38, 42, 45, 47, 64, 75, 89, 92, 96, 97, 102, 103, 110],
        [2, 3, 9, 10, 15, 18, 23, 24, 30, 31, 33, 37, 46, 81, 86, 90, 91, 93, 98, 102, 103, 104, 106, 112],
        [2, 3, 7, 12, 20, 23, 24, 30, 31, 34, 35, 36, 38, 42, 45, 47, 89, 92, 96, 98, 102, 103, 106, 111],
        [2, 3, 11, 12, 14, 16, 23, 24, 30, 31, 34, 36, 38, 47, 86, 87, 90, 95, 100, 101, 103, 111, 113],
        [5, 6, 7, 12, 19, 26, 27, 29, 34, 35, 36, 38, 79, 89, 91, 94, 97, 101, 103, 105, 106, 111, 112],
        [1, 6, 11, 12, 15, 18, 27, 28, 29, 34, 36, 38, 41, 52, 84, 90, 91, 94, 98, 102, 103, 106, 111],
        [3, 4, 7, 12, 20, 24, 25, 30, 31, 34, 35, 36, 38, 45, 48, 50, 88, 89, 95, 97, 101, 103, 106],
        [5, 6, 7, 12, 13, 16, 22, 26, 27, 29, 34, 35, 36, 38, 79, 89, 91, 94, 98, 99, 101, 103, 106],
        [4, 5, 7, 12, 14, 17, 25, 26, 34, 35, 36, 38, 48, 77, 89, 97, 100, 101, 103, 106, 107, 111],
        [2, 3, 7, 12, 18, 23, 24, 30, 31, 34, 35, 36, 38, 42, 45, 47, 56, 75, 89, 92, 96, 101, 103],
        [2, 3, 11, 12, 13, 21, 23, 24, 30, 31, 34, 36, 38, 47, 87, 90, 95, 99, 102, 103, 106, 111],
        [1, 6, 9, 10, 13, 16, 22, 27, 28, 29, 33, 37, 40, 69, 90, 95, 98, 99, 101, 102, 106, 111],
        [1, 6, 7, 12, 15, 18, 27, 28, 29, 34, 35, 36, 38, 39, 41, 64, 79, 89, 97, 101, 103, 109],
        [1, 2, 11, 12, 15, 18, 23, 28, 34, 36, 38, 44, 54, 90, 93, 97, 101, 103, 104, 106, 111],
        [3, 4, 11, 12, 13, 21, 24, 25, 30, 31, 34, 36, 38, 50, 57, 90, 96, 99, 101, 103, 111],
        [1, 6, 8, 9, 14, 16, 27, 28, 29, 32, 37, 68, 76, 89, 92, 96, 100, 101, 102, 110, 112],
        [5, 6, 8, 9, 13, 16, 22, 26, 27, 29, 32, 37, 80, 89, 95, 98, 99, 102, 103, 106, 111],
        [1, 6, 10, 11, 14, 17, 27, 28, 29, 33, 40, 51, 84, 89, 98, 100, 101, 103, 106, 111],
        [3, 4, 10, 11, 13, 16, 22, 24, 25, 30, 31, 33, 49, 89, 97, 99, 102, 103, 106, 108],
        [1, 2, 10, 11, 13, 16, 22, 23, 28, 33, 43, 72, 89, 91, 95, 99, 101, 103, 111, 113],
        [3, 4, 11, 12, 20, 24, 25, 30, 31, 34, 36, 38, 50, 74, 90, 96, 97, 102, 103, 111],
        [4, 5, 8, 9, 15, 18, 25, 26, 32, 37, 60, 67, 82, 89, 96, 98, 101, 103, 106, 112],
        [2, 3, 10, 11, 18, 23, 24, 30, 31, 33, 46, 56, 71, 86, 89, 96, 101, 102, 112],
        [4, 5, 10, 11, 15, 18, 25, 26, 33, 83, 89, 97, 102, 103, 106, 109, 112],
        [4, 5, 11, 12, 14, 16, 25, 26, 34, 36, 38, 59, 90, 100, 102, 103, 106],
        [3, 4, 7, 8, 14, 17, 24, 25, 30, 31, 32, 35, 45, 48, 65, 77, 85, 88, 90, 96, 97, 100, 101, 103, 105, 106, 107,
         112],
        [3, 4, 7, 12, 15, 17, 24, 25, 30, 31, 34, 35, 36, 38, 45, 48, 50, 63, 77, 85, 89, 95, 101, 103, 110, 112],
        [2, 3, 8, 9, 13, 16, 22, 23, 24, 30, 31, 32, 37, 55, 78, 89, 97, 99, 102, 103, 105, 106, 109, 112],
        [1, 2, 9, 10, 13, 16, 22, 23, 28, 33, 37, 43, 69, 70, 90, 96, 97, 99, 102, 103, 106, 108, 111],
        [5, 6, 7, 12, 14, 17, 26, 27, 29, 34, 35, 36, 38, 62, 79, 89, 91, 94, 97, 100, 101, 102, 107],
        [1, 2, 9, 10, 14, 16, 23, 28, 33, 37, 43, 53, 90, 96, 100, 101, 103, 105, 106, 112],
        [3, 4, 7, 8, 18, 24, 25, 30, 31, 32, 35, 45, 48, 66, 77, 85, 90, 96, 102, 103, 110],
        [4, 5, 11, 12, 19, 25, 26, 34, 36, 38, 73, 83, 90, 98, 101, 102, 106, 112, 113],
        [3, 4, 10, 11, 15, 18, 24, 25, 30, 31, 33, 49, 58, 89, 97, 101, 102, 106, 112],
        [5, 6, 11, 12, 13, 21, 26, 27, 29, 34, 36, 38, 61, 90, 99, 102, 103, 109]
    ],
}


# ===  ===
def get_bounds():
    mins = np.array([EXPERIMENT_CONFIG['MIN_X'], EXPERIMENT_CONFIG['MIN_Y'], EXPERIMENT_CONFIG['MIN_Z']])
    maxs = np.array([EXPERIMENT_CONFIG['MAX_X'], EXPERIMENT_CONFIG['MAX_Y'], EXPERIMENT_CONFIG['MAX_Z']])
    return mins, maxs

def clip_state(state):
    mins, maxs = get_bounds()
    return np.clip(state, mins, maxs)

def denormalize_state(normalized_state):
    """"""
    mins, maxs = get_bounds()
    return normalized_state * (maxs - mins) / 2 + (mins + maxs) / 2


def coverage_similarity(triggered, target_path):
    """
    Similarity: / target paths
    """
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0

    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)


def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']

    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']

    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']

    return reward


# ===   ===
def execute_Tr(weather, time_period, z):
    """执行验证规则并返回触发的分支"""
    triggered = set()

    # Fixed all if statements - using triggered.add() instead of b[0]=1
    if (weather == 1) != (weather == 2):
        triggered.add(1)
    if (weather == 2) != (weather == 3):
        triggered.add(2)
    if (weather == 3) != (weather == 4):
        triggered.add(3)
    if (weather == 4) != (weather == 5):
        triggered.add(4)
    if (weather == 5) != (weather == 6):
        triggered.add(5)
    if (weather == 6) != (weather == 1):
        triggered.add(6)

    # 时间段相关规则 (7-12)
    if (time_period == 1) != (time_period == 2):
        triggered.add(7)
    if (time_period == 2) != (time_period == 3):
        triggered.add(8)
    if (time_period == 3) != (time_period == 4):
        triggered.add(9)
    if (time_period == 4) != (time_period == 5):
        triggered.add(10)
    if (time_period == 5) != (time_period == 6):
        triggered.add(11)
    if (time_period == 6) != (time_period == 1):
        triggered.add(12)

    # 行人数量相关规则 (13-22)
    if (z < 20) != (z < 30):
        triggered.add(13)
    if (z < 30) != (z < 40):
        triggered.add(14)
    if (z < 40) != (z < 50):
        triggered.add(15)
    if (z > 20) != (z > 30):
        triggered.add(16)
    if (z > 30) != (z > 40):
        triggered.add(17)
    if (z > 40) != (z > 50):
        triggered.add(18)
    if (z > 50) != (z > 60):
        triggered.add(19)
    if (10 < z < 50) != (15 < z < 50):
        triggered.add(20)
    if (15 < z < 45) != (20 < z < 45):
        triggered.add(21)
    if (20 < z < 40) != (25 < z < 40):
        triggered.add(22)

    # 天气组合规则 (23-31)
    if (weather in [1, 2]) != (weather in [1, 3]):
        triggered.add(23)
    if (weather in [2, 3]) != (weather in [2, 4]):
        triggered.add(24)
    if (weather in [3, 4]) != (weather in [3, 5]):
        triggered.add(25)
    if (weather in [4, 5]) != (weather in [4, 6]):
        triggered.add(26)
    if (weather in [5, 6]) != (weather in [5, 1]):
        triggered.add(27)
    if (weather in [1, 3, 5]) != (weather in [2, 3, 5]):
        triggered.add(28)
    if (weather in [2, 4, 6]) != (weather in [2, 4, 1]):
        triggered.add(29)
    if (weather in [1, 2, 3]) != (weather in [1, 2, 4]):
        triggered.add(30)
    if (weather in [4, 5, 6]) != (weather in [3, 5, 6]):
        triggered.add(31)

    # 时间段组合规则 (32-38)
    if (time_period in [1, 2]) != (time_period in [1, 3]):
        triggered.add(32)
    if (time_period in [3, 4]) != (time_period in [3, 5]):
        triggered.add(33)
    if (time_period in [5, 6]) != (time_period in [5, 1]):
        triggered.add(34)
    if (time_period in [1, 3, 5]) != (time_period in [2, 3, 5]):
        triggered.add(35)
    if (time_period in [2, 4, 6]) != (time_period in [2, 4, 1]):
        triggered.add(36)
    if (time_period in [1, 2, 3]) != (time_period in [1, 2, 4]):
        triggered.add(37)
    if (time_period in [4, 5, 6]) != (time_period in [4, 5, 1]):
        triggered.add(38)

    # 天气和时间段交互规则 (39-50)
    if (weather == 1 and time_period in [1, 2]) != (weather == 2 and time_period in [1, 2]):
        triggered.add(39)
    if (weather == 1 and time_period in [3, 4]) != (weather == 1 and time_period in [3, 5]):
        triggered.add(40)
    if (weather == 1 and time_period in [5, 6]) != (weather == 1 and time_period in [5, 1]):
        triggered.add(41)
    if (weather == 2 and time_period in [1, 2]) != (weather == 3 and time_period in [1, 2]):
        triggered.add(42)
    if (weather == 2 and time_period in [3, 4]) != (weather == 2 and time_period in [3, 5]):
        triggered.add(43)
    if (weather == 2 and time_period in [5, 6]) != (weather == 2 and time_period in [5, 1]):
        triggered.add(44)
    if (weather == 3 and time_period in [1, 2]) != (weather == 4 and time_period in [1, 2]):
        triggered.add(45)
    if (weather == 3 and time_period in [3, 4]) != (weather == 3 and time_period in [3, 5]):
        triggered.add(46)
    if (weather == 3 and time_period in [5, 6]) != (weather == 3 and time_period in [5, 1]):
        triggered.add(47)
    if (weather == 4 and time_period in [1, 2]) != (weather == 5 and time_period in [1, 2]):
        triggered.add(48)
    if (weather == 4 and time_period in [3, 4]) != (weather == 4 and time_period in [3, 5]):
        triggered.add(49)
    if (weather == 4 and time_period in [5, 6]) != (weather == 4 and time_period in [5, 1]):
        triggered.add(50)

    # 天气和行人数量交互规则 (51-62)
    if (weather == 1 and z > 30) != (weather == 1 and z > 35):
        triggered.add(51)
    if (weather == 1 and z < 40) != (weather == 1 and z < 45):
        triggered.add(52)
    if (weather == 2 and z > 25) != (weather == 2 and z > 30):
        triggered.add(53)
    if (weather == 2 and z < 45) != (weather == 2 and z < 50):
        triggered.add(54)
    if (weather == 3 and z > 20) != (weather == 3 and z > 25):
        triggered.add(55)
    if (weather == 3 and z < 50) != (weather == 3 and z < 55):
        triggered.add(56)
    if (weather == 4 and z > 15) != (weather == 4 and z > 20):
        triggered.add(57)
    if (weather == 4 and z < 45) != (weather == 4 and z < 50):
        triggered.add(58)
    if (weather == 5 and z > 25) != (weather == 5 and z > 30):
        triggered.add(59)
    if (weather == 5 and z < 40) != (weather == 5 and z < 45):
        triggered.add(60)
    if (weather == 6 and z > 15) != (weather == 6 and z > 20):
        triggered.add(61)
    if (weather == 6 and z < 35) != (weather == 6 and z < 40):
        triggered.add(62)

    # 时间段和行人数量交互规则 (63-74)
    if (time_period == 1 and z > 35) != (time_period == 1 and z > 40):
        triggered.add(63)
    if (time_period == 1 and z < 45) != (time_period == 1 and z < 50):
        triggered.add(64)
    if (time_period == 2 and z > 30) != (time_period == 2 and z > 35):
        triggered.add(65)
    if (time_period == 2 and z < 50) != (time_period == 2 and z < 55):
        triggered.add(66)
    if (time_period == 3 and z > 40) != (time_period == 3 and z > 45):
        triggered.add(67)
    if (time_period == 3 and z < 35) != (time_period == 3 and z < 30):
        triggered.add(68)
    if (time_period == 4 and z > 20) != (time_period == 4 and z > 25):
        triggered.add(69)
    if (time_period == 4 and z < 30) != (time_period == 4 and z < 25):
        triggered.add(70)
    if (time_period == 5 and z > 45) != (time_period == 5 and z > 50):
        triggered.add(71)
    if (time_period == 5 and z < 25) != (time_period == 5 and z < 20):
        triggered.add(72)
    if (time_period == 6 and z > 50) != (time_period == 6 and z > 55):
        triggered.add(73)
    if (time_period == 6 and z < 20) != (time_period == 6 and z < 15):
        triggered.add(74)

    # 三元素组合规则 (75-84)
    if (weather in [1, 2] and time_period in [1, 2] and z > 30) != (
            weather in [1, 3] and time_period in [1, 2] and z > 30):
        triggered.add(75)
    if (weather in [1, 2] and time_period in [1, 2] and z < 40) != (
            weather in [1, 2] and time_period in [1, 3] and z < 40):
        triggered.add(76)
    if (weather in [3, 4] and time_period in [1, 2] and z > 25) != (
            weather in [3, 5] and time_period in [1, 2] and z > 25):
        triggered.add(77)
    if (weather in [3, 4] and time_period in [1, 2] and z < 35) != (
            weather in [3, 4] and time_period in [1, 3] and z < 35):
        triggered.add(78)
    if (weather in [5, 6] and time_period in [1, 2] and z > 20) != (
            weather in [5, 1] and time_period in [1, 2] and z > 20):
        triggered.add(79)
    if (weather in [5, 6] and time_period in [1, 2] and z < 30) != (
            weather in [5, 6] and time_period in [1, 3] and z < 30):
        triggered.add(80)
    if (weather in [1, 3] and time_period in [3, 4] and z > 35) != (
            weather in [1, 4] and time_period in [3, 4] and z > 35):
        triggered.add(81)
    if (weather in [2, 4] and time_period in [3, 4] and z > 30) != (
            weather in [2, 5] and time_period in [3, 4] and z > 30):
        triggered.add(82)
    if (weather in [1, 5] and time_period in [5, 6] and z > 40) != (
            weather in [1, 6] and time_period in [5, 6] and z > 40):
        triggered.add(83)
    if (weather in [2, 6] and time_period in [5, 6] and z > 25) != (
            weather in [2, 1] and time_period in [5, 6] and z > 25):
        triggered.add(84)

    # 复杂条件规则 (85-100)
    if (weather <= 3 and time_period <= 3 and z > 25) != (weather <= 4 and time_period <= 3 and z > 25):
        triggered.add(85)
    if (weather >= 4 and time_period >= 4 and z > 20) != (weather >= 3 and time_period >= 4 and z > 20):
        triggered.add(86)
    if (weather <= 2 and time_period >= 4 and z < 35) != (weather <= 3 and time_period >= 4 and z < 35):
        triggered.add(87)
    if (weather >= 5 and time_period <= 2 and z < 40) != (weather >= 4 and time_period <= 2 and z < 40):
        triggered.add(88)
    if (weather % 2 == 1 and time_period % 2 == 1) != (weather % 2 == 0 and time_period % 2 == 1):
        triggered.add(89)
    if (weather % 2 == 0 and time_period % 2 == 0) != (weather % 2 == 1 and time_period % 2 == 0):
        triggered.add(90)
    if (weather + time_period > 6) != (weather + time_period > 7):
        triggered.add(91)
    if (weather + time_period < 5) != (weather + time_period < 4):
        triggered.add(92)
    if (weather * time_period > 10) != (weather * time_period > 12):
        triggered.add(93)
    if (weather * time_period < 8) != (weather * time_period < 6):
        triggered.add(94)
    if (abs(weather - time_period) <= 2) != (abs(weather - time_period) <= 3):
        triggered.add(95)
    if (abs(weather - time_period) >= 3) != (abs(weather - time_period) >= 2):
        triggered.add(96)
    if (z % 10 < 5) != (z % 10 < 6):
        triggered.add(97)
    if (z % 10 >= 5) != (z % 10 >= 4):
        triggered.add(98)
    if (z // 10 >= 3) != (z // 10 >= 2):
        triggered.add(99)
    if (z // 10 <= 2) != (z // 10 <= 3):
        triggered.add(100)

    # 高级组合规则 (101-113)
    if ((weather + time_period + z // 10) % 3 == 0) != ((weather + time_period + z // 10) % 3 == 1):
        triggered.add(101)
    if ((weather + time_period + z // 10) % 3 == 1) != ((weather + time_period + z // 10) % 3 == 2):
        triggered.add(102)
    if ((weather + time_period + z // 10) % 3 == 2) != ((weather + time_period + z // 10) % 3 == 0):
        triggered.add(103)
    if (weather * time_period + z // 10 > 15) != (weather * time_period + z // 10 > 16):
        triggered.add(104)
    if (weather * time_period + z // 10 < 12) != (weather * time_period + z // 10 < 11):
        triggered.add(105)
    if ((weather * time_period) % (z // 10 + 1) == 0) != ((weather * time_period) % (z // 10 + 2) == 0):
        triggered.add(106)
    if (weather > time_period and z > 30) != (weather > time_period and z > 35):
        triggered.add(107)
    if (weather < time_period and z < 30) != (weather < time_period and z < 25):
        triggered.add(108)
    if (weather == time_period) != (weather == time_period + 1):
        triggered.add(109)
    if (weather + time_period == z // 10) != (weather + time_period == z // 10 + 1):
        triggered.add(110)
    if (abs(weather - time_period) == z // 10) != (abs(weather - time_period) == z // 10 + 1):
        triggered.add(111)
    if (max(weather, time_period) == z // 10) != (max(weather, time_period) == z // 10 + 1):
        triggered.add(112)
    if (min(weather, time_period) * 10 <= z) != (min(weather, time_period) * 11 <= z):
        triggered.add(113)

    return triggered

# 
execute_Tr = execute_Tr

# === DQN ===
class DQNNetwork(nn.Module):
    def __init__(self, action_size=30):
        super(DQNNetwork, self).__init__()
        hidden_dim = EXPERIMENT_CONFIG['HIDDEN_DIM']

        self.fc1 = nn.Linear(3, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)
        self.output = nn.Linear(hidden_dim, action_size)

    def forward(self, x):
        x = torch.relu(self.fc1(x))
        x = torch.relu(self.fc2(x))
        x = torch.relu(self.fc3(x))
        return self.output(x)


# === Path  ===
class PathReplayBuffer:
    """Path """

    def __init__(self, path_idx, capacity=20000):
        self.path_idx = path_idx
        self.capacity = capacity
        self.buffer = deque(maxlen=capacity)
        self.similarities = deque(maxlen=capacity)  # Similarity

    def push(self, state, action, reward, next_state, done, similarity):
        self.buffer.append((state, action, reward, next_state, done))
        self.similarities.append(similarity)

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return None

        batch = random.sample(self.buffer, batch_size)
        state, action, reward, next_state, done = map(np.stack, zip(*batch))

        return (
            torch.FloatTensor(state).to(device),
            torch.LongTensor(action).to(device),
            torch.FloatTensor(reward).to(device),
            torch.FloatTensor(next_state).to(device),
            torch.BoolTensor(done).to(device)
        )

    def get_top_k(self, k=20):
        """Path Top-K"""
        if len(self.buffer) == 0:
            return []

        # buffersimilarities
        samples_with_sim = list(zip(self.buffer, self.similarities))
        samples_with_sim.sort(key=lambda x: x[1], reverse=True)

        # Top-K
        top_k = samples_with_sim[:k]

        results = []
        for (state, _, _, _, _), similarity in top_k:
            original_state = denormalize_state(state)
            original_state_int = np.round(original_state).astype(int)

            results.append({
                'state': original_state_int,
                'similarity': similarity,
                'triggered': execute_Tr(*original_state_int)
            })

        return results

    def __len__(self):
        return len(self.buffer)


# === DQN(:)===
class ImprovedDQNAgent:
    def __init__(self, num_paths, action_size=30):
        self.action_size = action_size
        self.num_paths = num_paths
        self.epsilon = 0.9
        self.epsilon_min = 0.1
        self.epsilon_decay = 0.995

        self.q_network = DQNNetwork(action_size).to(device)
        self.target_network = DQNNetwork(action_size).to(device)

        lr = EXPERIMENT_CONFIG['LEARNING_RATE']
        self.optimizer = optim.Adam(self.q_network.parameters(), lr=lr)

        # Path 
        capacity = EXPERIMENT_CONFIG['REPLAY_BUFFER_CAPACITY']
        self.replay_buffers = {}
        for path_idx in range(num_paths):
            self.replay_buffers[path_idx] = PathReplayBuffer(path_idx, capacity)

        self.replay_train_count = 0
        self.update_target_network()

    def discrete_to_action_delta(self, action_idx):
        # 
        delta_values = [5, 3, 2, 1, 0.5, -0.5, -1, -2, -3, -5]

        if action_idx >= 30:
            action_idx = action_idx % 30

        dim = action_idx // 10
        delta_idx = action_idx % 10
        delta = delta_values[delta_idx]

        action_delta = np.zeros(3)
        action_delta[dim] = delta

        return action_delta

    def get_action(self, state):
        mins, maxs = get_bounds()
        normalized_state = (state - (mins + maxs) / 2) / ((maxs - mins) / 2)

        if random.random() < self.epsilon:
            action_idx = random.randint(0, self.action_size - 1)
        else:
            state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)
            with torch.no_grad():
                q_values = self.q_network(state_tensor)
                action_idx = q_values.argmax().item()

        action_delta = self.discrete_to_action_delta(action_idx)
        return action_delta, action_idx

    def store_experience(self, path_idx, state, action_idx, reward, next_state, done, similarity):
        """Path """
        mins, maxs = get_bounds()
        normalized_state = (state - (mins + maxs) / 2) / ((maxs - mins) / 2)
        normalized_next_state = (next_state - (mins + maxs) / 2) / ((maxs - mins) / 2)

        self.replay_buffers[path_idx].push(
            normalized_state, action_idx, reward,
            normalized_next_state, done, similarity
        )

    def replay_train(self, path_idx):
        """Path """
        batch_size = EXPERIMENT_CONFIG['REPLAY_BATCH_SIZE']
        batch = self.replay_buffers[path_idx].sample(batch_size)

        if batch is None:
            return

        states, actions, rewards, next_states, dones = batch

        current_q_values = self.q_network(states).gather(1, actions.unsqueeze(1))

        with torch.no_grad():
            next_q_values = self.target_network(next_states).max(1)[0]
            target_q_values = rewards + (0.99 * next_q_values * ~dones)

        loss = F.mse_loss(current_q_values.squeeze(), target_q_values)

        self.optimizer.zero_grad()
        loss.backward()
        torch.nn.utils.clip_grad_norm_(self.q_network.parameters(), 1.0)
        self.optimizer.step()

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

        self.replay_train_count += 1

        if self.replay_train_count % 2 == 0:
            self.update_target_network()
            print(f"    ->  (Run {self.replay_train_count})")

    def update_target_network(self):
        self.target_network.load_state_dict(self.q_network.state_dict())

    def get_all_top_k(self, k=20):
        """Path Top-K"""
        results = {}
        for path_idx in range(self.num_paths):
            results[path_idx] = self.replay_buffers[path_idx].get_top_k(k)
        return results

    def get_buffer_stats(self):
        """"""
        stats = {}
        for path_idx in range(self.num_paths):
            stats[path_idx] = len(self.replay_buffers[path_idx])
        return stats


# === Metric ===
def calculate_run_performance(run_idx, dqn_results, training_time, total_steps, update_count, agent):
    """ runMetric"""
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    # 1. (Total Reward)
    total_reward = 0
    # 2. (Average Reward)
    average_reward = 0
    # 5. (Convergence)
    convergence = 0
    # 12. (Environment Adaptability)
    environment_adaptability = 0
    # 13. (Generalization Ability)
    generalization_ability = 0
    # 15. (Computational Efficiency)
    computational_efficiency = 0
    # 16. (Policy Update Frequency)
    policy_update_frequency = 0

    # Similarity
    all_similarities = []

    # Metric
    total_samples = 0
    all_rewards = []

    for path_idx in range(num_paths):
        samples = dqn_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']

            total_reward += reward
            all_rewards.append(reward)
            all_similarities.append(similarity)
            total_samples += 1

    # 1. 
    total_reward = total_reward

    # 2. 
    if total_samples > 0:
        average_reward = total_reward / total_samples

    # 5. (Average Similarity)
    if all_similarities:
        convergence = np.mean(all_similarities)

    # 12. (Similarity)
    if len(all_similarities) > 1:
        environment_adaptability = 1 / (np.std(all_similarities) + 1e-8)

    # 13. (Average Similarity)
    generalization_ability = convergence

    # 15. (/ seconds)
    if training_time > 0:
        computational_efficiency = total_steps / training_time

    # 16. 
    if training_time > 0:
        policy_update_frequency = update_count / training_time

    # Similarity
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0

    return {
        'Run': run_idx + 1,

        # Metric
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),

        # Similarity
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }


# === Excel ===
def export_to_excel(all_dqn_results, all_performance_data, target_paths, output_path="DQN_20 run.xlsx"):
    """20 runDQNExcel"""
    print("\nExcel...")

    # 
    all_dqn_summary_data = []
    all_dqn_detailed_data = []

    #  run
    for run_idx, (dqn_results, performance_data) in enumerate(zip(all_dqn_results, all_performance_data)):
        # ===== Sheet1: DQNPath  =====
        dqn_summary_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = dqn_results[path_idx]

            if len(samples) == 0:
                dqn_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue

            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = 'Yes' if perfect_count > 0 else 'No'

            dqn_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })
        all_dqn_summary_data.extend(dqn_summary_data)

        # ===== Sheet2: DQNDetailed Sample Data =====
        dqn_detailed_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = dqn_results[path_idx]

            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']

                dqn_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Sample ID': sample_idx + 1,
                    'X': int(state[0]),
                    'Y': int(state[1]),
                    'Z': int(state[2]),
                    'Similarity': round(similarity, 4),
                    'Perfect Coverage': 'Yes' if abs(similarity - 1.0) < 0.001 else 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path))),
                    'Triggered Rules': ', '.join(map(str, sorted(triggered))),
                    'Hit Rule Count': len(target_path.intersection(triggered)),
                    'Target Rule Count': len(target_path)
                })
        all_dqn_detailed_data.extend(dqn_detailed_data)

    # Excel
    dqn_summary_df = pd.DataFrame(all_dqn_summary_data)
    dqn_detailed_df = pd.DataFrame(all_dqn_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet1: DQNPath 
        dqn_summary_df.to_excel(writer, sheet_name='DQNPath ', index=False)

        # Sheet2: DQNDetailed Sample Data
        dqn_detailed_df.to_excel(writer, sheet_name='DQNDetailed Sample Data', index=False)

        # Sheet3: Metric
        selected_columns = [
            'Run',
            'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df_selected = performance_df[selected_columns]
        performance_df_selected.to_excel(writer, sheet_name='Metric', index=False)

        # 
        workbook = writer.book

        # 
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        perfect_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # 

        # === Sheet1 ===
        ws1 = writer.sheets['DQNPath ']
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 
        for row_idx in range(2, ws1.max_row + 1):
            if ws1.cell(row_idx, 9).value == '':  # Run 9""
                for col_idx in range(1, ws1.max_column + 1):
                    ws1.cell(row_idx, col_idx).fill = perfect_fill

        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 15
        ws1.column_dimensions['F'].width = 15
        ws1.column_dimensions['G'].width = 15
        ws1.column_dimensions['H'].width = 15
        ws1.column_dimensions['I'].width = 15
        ws1.column_dimensions['J'].width = 50

        # === Sheet2 ===
        ws2 = writer.sheets['DQNDetailed Sample Data']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 15
        ws2.column_dimensions['I'].width = 40
        ws2.column_dimensions['J'].width = 40
        ws2.column_dimensions['K'].width = 15
        ws2.column_dimensions['L'].width = 15

        # === Sheet3 ===
        ws3 = writer.sheets['Metric']
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        for col in columns:
            ws3.column_dimensions[col].width = 18

    print(f"Excel: {output_path}")
    print(f"  - Sheet1: DQNPath  ({len(all_dqn_summary_data)})")
    print(f"  - Sheet2: DQNDetailed Sample Data ({len(all_dqn_detailed_data)})")
    print(f"  - Sheet3: Metric ({len(all_performance_data)})")


# === DQN training(:)===
def train_dqn_workflow():
    print("=" * 80)
    print("DQN training ()")
    print("Similarity:  / target paths")
    print(
        f": Path {EXPERIMENT_CONFIG['NUM_ROUNDS']},{EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(f": Path ,={EXPERIMENT_CONFIG['REPLAY_BUFFER_CAPACITY']}")
    print("=" * 80)

    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    # (Number of Paths)
    agent = ImprovedDQNAgent(num_paths=num_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n: Path {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']}")
    path_samples = {}
    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            state = np.array([
                random.randint(EXPERIMENT_CONFIG['MIN_X'], EXPERIMENT_CONFIG['MAX_X']),
                random.randint(EXPERIMENT_CONFIG['MIN_Y'], EXPERIMENT_CONFIG['MAX_Y']),
                random.randint(EXPERIMENT_CONFIG['MIN_Z'], EXPERIMENT_CONFIG['MAX_Z'])
            ], dtype=np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  Path  {path_idx + 1}/{num_paths}:  {len(samples)} ")

    # 
    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size
    num_rounds = EXPERIMENT_CONFIG['NUM_ROUNDS']

    print(f"\n:")
    print(f"  - : {batch_size}")
    print(f"  - Path : {num_batches}")
    print(f"  - Path : {num_rounds}")
    print(f"  - : {EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(
        f"  - : {num_paths} Path  x {num_rounds}  x {num_batches}  = {num_paths * num_rounds * num_batches} ")
    print("-" * 80)

    # :completedPath ,Path 
    for path_idx in range(num_paths):
        target_path = target_paths[path_idx]
        print(f"\n{'=' * 80}")
        print(f"Start training path  {path_idx + 1}/{num_paths}")
        print(f": {sorted(target_path)}")
        print(f": replay_buffers[{path_idx}]")
        print(f"{'=' * 80}")

        # Path NUM_ROUNDS
        for round_idx in range(num_rounds):
            print(f"\n{'' * 80}")
            print(f"Path  {path_idx + 1} - Run  {round_idx + 1}/{num_rounds} ")
            print(f"{'' * 80}")

            # Per roundnum_batches
            for batch_idx in range(num_batches):
                print(f"\n   {batch_idx + 1}/{num_batches} (Path {path_idx + 1}, Run {round_idx + 1})")

                # 
                batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]

                batch_rewards = []
                batch_similarities = []

                # 
                for sample_idx, initial_state in enumerate(batch_samples):
                    state = initial_state.copy()
                    episode_reward = 0
                    final_similarity = 0

                    # STEPS_PER_SAMPLE
                    for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                        action_delta, action_idx = agent.get_action(state)

                        next_state = state + action_delta
                        next_state = clip_state(next_state)

                        triggered = execute_Tr(*next_state)  #
                        reward = unified_reward_function(triggered, target_path)
                        similarity = coverage_similarity(triggered, target_path)

                        done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)

                        # Path 
                        agent.store_experience(
                            path_idx, state, action_idx, reward, next_state, done, similarity
                        )

                        state = next_state
                        episode_reward += reward
                        final_similarity = similarity
                        total_steps += 1

                    batch_rewards.append(episode_reward)
                    batch_similarities.append(final_similarity)

                # 
                avg_reward = np.mean(batch_rewards)
                avg_similarity = np.mean(batch_similarities)
                max_similarity = np.max(batch_similarities)
                print(f"    ={avg_reward:.2f}, Average Similarity={avg_similarity:.4f}, "
                      f"Similarity={max_similarity:.4f}, epsilon={agent.epsilon:.3f}")

                # Path 
                print(f"    (Path {path_idx})...")
                agent.replay_train(path_idx)

                # Path 
                buffer_size = len(agent.replay_buffers[path_idx])
                print(f"    Path {path_idx}: {buffer_size}, : {agent.replay_train_count}")

    training_time = time.time() - start_time

    print("\n" + "=" * 80)
    print(f"DQN trainingcompleted! Total elapsed time: {training_time:.2f} seconds, : {total_steps}")
    print(f": {agent.replay_train_count}")
    print(f": {agent.replay_train_count // 2}")

    # Path 
    print("\nPath :")
    buffer_stats = agent.get_buffer_stats()
    for path_idx, size in buffer_stats.items():
        print(f"  Path {path_idx + 1}: {size} ")

    print("=" * 80)

    # Top-K
    print(f"\nPath SimilarityMaximum{EXPERIMENT_CONFIG['TOP_K_SAMPLES']}...")
    dqn_top_k_results = agent.get_all_top_k(EXPERIMENT_CONFIG['TOP_K_SAMPLES'])

    return agent, dqn_top_k_results, training_time, total_steps, agent.replay_train_count


# ===  ===
def main():
    print("\n" + "=" * 80)
    print("DQN - 20 run")
    print("Metric")
    print("=" * 80)

    all_dqn_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    # 20
    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'='*80}")
        print(f"Start run  {run_idx + 1}/{EXPERIMENT_CONFIG['NUM_RUNS']}  run")
        print(f"{'='*80}")

        # DQN training
        dqn_agent, dqn_results, training_time, total_steps, update_count = train_dqn_workflow()

        # Metric
        performance_data = calculate_run_performance(
            run_idx, dqn_results, training_time, total_steps, update_count, dqn_agent
        )

        # 
        all_dqn_results.append(dqn_results)
        all_performance_data.append(performance_data)

        print(f"\nRun {run_idx + 1} completed!")
        print(f"  Total Reward: {performance_data['Total Reward']}")
        print(f"  Average Reward: {performance_data['Average Reward']}")
        print(f"  Convergence: {performance_data['Convergence']}")

    # Excel(20 run)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"DQN_20 run_{timestamp}.xlsx"
    export_to_excel(all_dqn_results, all_performance_data, target_paths, output_path)

    # 
    print("\n" + "=" * 80)
    print("20 run")
    print("=" * 80)

    # Metric
    # Metric Extraction
    total_rewards = [p['Total Reward'] for p in all_performance_data]
    average_rewards = [p['Average Reward'] for p in all_performance_data]
    convergences = [p['Convergence'] for p in all_performance_data]
    environment_adaptabilities = [p['Environment Adaptability'] for p in all_performance_data]
    generalization_abilities = [p['Generalization Ability'] for p in all_performance_data]
    computational_efficiencies = [p['Computational Efficiency'] for p in all_performance_data]
    policy_update_frequencies = [p['Policy Update Frequency'] for p in all_performance_data]
    avg_similarities = [p['Average Similarity'] for p in all_performance_data]

    print(f":")
    print(f"  : {np.mean(total_rewards):.2f}")
    print(f"  Standard deviation: {np.std(total_rewards):.2f}")

    print(f"\n:")
    print(f"  : {np.mean(average_rewards):.4f}")
    print(f"  Standard deviation: {np.std(average_rewards):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(convergences):.4f}")
    print(f"  Standard deviation: {np.std(convergences):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(environment_adaptabilities):.4f}")
    print(f"  Standard deviation: {np.std(environment_adaptabilities):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(generalization_abilities):.4f}")
    print(f"  Standard deviation: {np.std(generalization_abilities):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(computational_efficiencies):.2f}")
    print(f"  Standard deviation: {np.std(computational_efficiencies):.2f}")

    print(f"\n:")
    print(f"  : {np.mean(policy_update_frequencies):.4f}")
    print(f"  Standard deviation: {np.std(policy_update_frequencies):.4f}")

    print(f"\nAverage similarity statistics:")
    print(f"  : {np.mean(avg_similarities):.4f}")
    print(f"  Standard deviation: {np.std(avg_similarities):.4f}")

    print("\n" + "=" * 80)
    print(f" {EXPERIMENT_CONFIG['NUM_RUNS']} completed!")
    print("=" * 80)


if __name__ == "__main__":
    main()