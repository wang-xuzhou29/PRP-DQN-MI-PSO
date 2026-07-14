import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

def execute_validation_rules(weather, time_period, z):
    """执行验证规则并返回触发的分支"""
    triggered = set()

    # Fixed all if statements - using triggered.add() instead of b[0]=1
    if (weather == 1) != (weather == 2):
        triggered.add(1)
    if (weather == 2) != (weather == 3):
        triggered.add(2)
    if (weather == 3) != (weather == 4):
        triggered.add(3)
    if (weather == 4) != (weather == 5):
        triggered.add(4)
    if (weather == 5) != (weather == 6):
        triggered.add(5)
    if (weather == 6) != (weather == 1):
        triggered.add(6)

    # 时间段相关规则 (7-12)
    if (time_period == 1) != (time_period == 2):
        triggered.add(7)
    if (time_period == 2) != (time_period == 3):
        triggered.add(8)
    if (time_period == 3) != (time_period == 4):
        triggered.add(9)
    if (time_period == 4) != (time_period == 5):
        triggered.add(10)
    if (time_period == 5) != (time_period == 6):
        triggered.add(11)
    if (time_period == 6) != (time_period == 1):
        triggered.add(12)

    # 行人数量相关规则 (13-22)
    if (z < 20) != (z < 30):
        triggered.add(13)
    if (z < 30) != (z < 40):
        triggered.add(14)
    if (z < 40) != (z < 50):
        triggered.add(15)
    if (z > 20) != (z > 30):
        triggered.add(16)
    if (z > 30) != (z > 40):
        triggered.add(17)
    if (z > 40) != (z > 50):
        triggered.add(18)
    if (z > 50) != (z > 60):
        triggered.add(19)
    if (10 < z < 50) != (15 < z < 50):
        triggered.add(20)
    if (15 < z < 45) != (20 < z < 45):
        triggered.add(21)
    if (20 < z < 40) != (25 < z < 40):
        triggered.add(22)

    # 天气组合规则 (23-31)
    if (weather in [1, 2]) != (weather in [1, 3]):
        triggered.add(23)
    if (weather in [2, 3]) != (weather in [2, 4]):
        triggered.add(24)
    if (weather in [3, 4]) != (weather in [3, 5]):
        triggered.add(25)
    if (weather in [4, 5]) != (weather in [4, 6]):
        triggered.add(26)
    if (weather in [5, 6]) != (weather in [5, 1]):
        triggered.add(27)
    if (weather in [1, 3, 5]) != (weather in [2, 3, 5]):
        triggered.add(28)
    if (weather in [2, 4, 6]) != (weather in [2, 4, 1]):
        triggered.add(29)
    if (weather in [1, 2, 3]) != (weather in [1, 2, 4]):
        triggered.add(30)
    if (weather in [4, 5, 6]) != (weather in [3, 5, 6]):
        triggered.add(31)

    # 时间段组合规则 (32-38)
    if (time_period in [1, 2]) != (time_period in [1, 3]):
        triggered.add(32)
    if (time_period in [3, 4]) != (time_period in [3, 5]):
        triggered.add(33)
    if (time_period in [5, 6]) != (time_period in [5, 1]):
        triggered.add(34)
    if (time_period in [1, 3, 5]) != (time_period in [2, 3, 5]):
        triggered.add(35)
    if (time_period in [2, 4, 6]) != (time_period in [2, 4, 1]):
        triggered.add(36)
    if (time_period in [1, 2, 3]) != (time_period in [1, 2, 4]):
        triggered.add(37)
    if (time_period in [4, 5, 6]) != (time_period in [4, 5, 1]):
        triggered.add(38)

    # 天气和时间段交互规则 (39-50)
    if (weather == 1 and time_period in [1, 2]) != (weather == 2 and time_period in [1, 2]):
        triggered.add(39)
    if (weather == 1 and time_period in [3, 4]) != (weather == 1 and time_period in [3, 5]):
        triggered.add(40)
    if (weather == 1 and time_period in [5, 6]) != (weather == 1 and time_period in [5, 1]):
        triggered.add(41)
    if (weather == 2 and time_period in [1, 2]) != (weather == 3 and time_period in [1, 2]):
        triggered.add(42)
    if (weather == 2 and time_period in [3, 4]) != (weather == 2 and time_period in [3, 5]):
        triggered.add(43)
    if (weather == 2 and time_period in [5, 6]) != (weather == 2 and time_period in [5, 1]):
        triggered.add(44)
    if (weather == 3 and time_period in [1, 2]) != (weather == 4 and time_period in [1, 2]):
        triggered.add(45)
    if (weather == 3 and time_period in [3, 4]) != (weather == 3 and time_period in [3, 5]):
        triggered.add(46)
    if (weather == 3 and time_period in [5, 6]) != (weather == 3 and time_period in [5, 1]):
        triggered.add(47)
    if (weather == 4 and time_period in [1, 2]) != (weather == 5 and time_period in [1, 2]):
        triggered.add(48)
    if (weather == 4 and time_period in [3, 4]) != (weather == 4 and time_period in [3, 5]):
        triggered.add(49)
    if (weather == 4 and time_period in [5, 6]) != (weather == 4 and time_period in [5, 1]):
        triggered.add(50)

    # 天气和行人数量交互规则 (51-62)
    if (weather == 1 and z > 30) != (weather == 1 and z > 35):
        triggered.add(51)
    if (weather == 1 and z < 40) != (weather == 1 and z < 45):
        triggered.add(52)
    if (weather == 2 and z > 25) != (weather == 2 and z > 30):
        triggered.add(53)
    if (weather == 2 and z < 45) != (weather == 2 and z < 50):
        triggered.add(54)
    if (weather == 3 and z > 20) != (weather == 3 and z > 25):
        triggered.add(55)
    if (weather == 3 and z < 50) != (weather == 3 and z < 55):
        triggered.add(56)
    if (weather == 4 and z > 15) != (weather == 4 and z > 20):
        triggered.add(57)
    if (weather == 4 and z < 45) != (weather == 4 and z < 50):
        triggered.add(58)
    if (weather == 5 and z > 25) != (weather == 5 and z > 30):
        triggered.add(59)
    if (weather == 5 and z < 40) != (weather == 5 and z < 45):
        triggered.add(60)
    if (weather == 6 and z > 15) != (weather == 6 and z > 20):
        triggered.add(61)
    if (weather == 6 and z < 35) != (weather == 6 and z < 40):
        triggered.add(62)

    # 时间段和行人数量交互规则 (63-74)
    if (time_period == 1 and z > 35) != (time_period == 1 and z > 40):
        triggered.add(63)
    if (time_period == 1 and z < 45) != (time_period == 1 and z < 50):
        triggered.add(64)
    if (time_period == 2 and z > 30) != (time_period == 2 and z > 35):
        triggered.add(65)
    if (time_period == 2 and z < 50) != (time_period == 2 and z < 55):
        triggered.add(66)
    if (time_period == 3 and z > 40) != (time_period == 3 and z > 45):
        triggered.add(67)
    if (time_period == 3 and z < 35) != (time_period == 3 and z < 30):
        triggered.add(68)
    if (time_period == 4 and z > 20) != (time_period == 4 and z > 25):
        triggered.add(69)
    if (time_period == 4 and z < 30) != (time_period == 4 and z < 25):
        triggered.add(70)
    if (time_period == 5 and z > 45) != (time_period == 5 and z > 50):
        triggered.add(71)
    if (time_period == 5 and z < 25) != (time_period == 5 and z < 20):
        triggered.add(72)
    if (time_period == 6 and z > 50) != (time_period == 6 and z > 55):
        triggered.add(73)
    if (time_period == 6 and z < 20) != (time_period == 6 and z < 15):
        triggered.add(74)

    # 三元素组合规则 (75-84)
    if (weather in [1, 2] and time_period in [1, 2] and z > 30) != (
            weather in [1, 3] and time_period in [1, 2] and z > 30):
        triggered.add(75)
    if (weather in [1, 2] and time_period in [1, 2] and z < 40) != (
            weather in [1, 2] and time_period in [1, 3] and z < 40):
        triggered.add(76)
    if (weather in [3, 4] and time_period in [1, 2] and z > 25) != (
            weather in [3, 5] and time_period in [1, 2] and z > 25):
        triggered.add(77)
    if (weather in [3, 4] and time_period in [1, 2] and z < 35) != (
            weather in [3, 4] and time_period in [1, 3] and z < 35):
        triggered.add(78)
    if (weather in [5, 6] and time_period in [1, 2] and z > 20) != (
            weather in [5, 1] and time_period in [1, 2] and z > 20):
        triggered.add(79)
    if (weather in [5, 6] and time_period in [1, 2] and z < 30) != (
            weather in [5, 6] and time_period in [1, 3] and z < 30):
        triggered.add(80)
    if (weather in [1, 3] and time_period in [3, 4] and z > 35) != (
            weather in [1, 4] and time_period in [3, 4] and z > 35):
        triggered.add(81)
    if (weather in [2, 4] and time_period in [3, 4] and z > 30) != (
            weather in [2, 5] and time_period in [3, 4] and z > 30):
        triggered.add(82)
    if (weather in [1, 5] and time_period in [5, 6] and z > 40) != (
            weather in [1, 6] and time_period in [5, 6] and z > 40):
        triggered.add(83)
    if (weather in [2, 6] and time_period in [5, 6] and z > 25) != (
            weather in [2, 1] and time_period in [5, 6] and z > 25):
        triggered.add(84)

    # 复杂条件规则 (85-100)
    if (weather <= 3 and time_period <= 3 and z > 25) != (weather <= 4 and time_period <= 3 and z > 25):
        triggered.add(85)
    if (weather >= 4 and time_period >= 4 and z > 20) != (weather >= 3 and time_period >= 4 and z > 20):
        triggered.add(86)
    if (weather <= 2 and time_period >= 4 and z < 35) != (weather <= 3 and time_period >= 4 and z < 35):
        triggered.add(87)
    if (weather >= 5 and time_period <= 2 and z < 40) != (weather >= 4 and time_period <= 2 and z < 40):
        triggered.add(88)
    if (weather % 2 == 1 and time_period % 2 == 1) != (weather % 2 == 0 and time_period % 2 == 1):
        triggered.add(89)
    if (weather % 2 == 0 and time_period % 2 == 0) != (weather % 2 == 1 and time_period % 2 == 0):
        triggered.add(90)
    if (weather + time_period > 6) != (weather + time_period > 7):
        triggered.add(91)
    if (weather + time_period < 5) != (weather + time_period < 4):
        triggered.add(92)
    if (weather * time_period > 10) != (weather * time_period > 12):
        triggered.add(93)
    if (weather * time_period < 8) != (weather * time_period < 6):
        triggered.add(94)
    if (abs(weather - time_period) <= 2) != (abs(weather - time_period) <= 3):
        triggered.add(95)
    if (abs(weather - time_period) >= 3) != (abs(weather - time_period) >= 2):
        triggered.add(96)
    if (z % 10 < 5) != (z % 10 < 6):
        triggered.add(97)
    if (z % 10 >= 5) != (z % 10 >= 4):
        triggered.add(98)
    if (z // 10 >= 3) != (z // 10 >= 2):
        triggered.add(99)
    if (z // 10 <= 2) != (z // 10 <= 3):
        triggered.add(100)

    # 高级组合规则 (101-113)
    if ((weather + time_period + z // 10) % 3 == 0) != ((weather + time_period + z // 10) % 3 == 1):
        triggered.add(101)
    if ((weather + time_period + z // 10) % 3 == 1) != ((weather + time_period + z // 10) % 3 == 2):
        triggered.add(102)
    if ((weather + time_period + z // 10) % 3 == 2) != ((weather + time_period + z // 10) % 3 == 0):
        triggered.add(103)
    if (weather * time_period + z // 10 > 15) != (weather * time_period + z // 10 > 16):
        triggered.add(104)
    if (weather * time_period + z // 10 < 12) != (weather * time_period + z // 10 < 11):
        triggered.add(105)
    if ((weather * time_period) % (z // 10 + 1) == 0) != ((weather * time_period) % (z // 10 + 2) == 0):
        triggered.add(106)
    if (weather > time_period and z > 30) != (weather > time_period and z > 35):
        triggered.add(107)
    if (weather < time_period and z < 30) != (weather < time_period and z < 25):
        triggered.add(108)
    if (weather == time_period) != (weather == time_period + 1):
        triggered.add(109)
    if (weather + time_period == z // 10) != (weather + time_period == z // 10 + 1):
        triggered.add(110)
    if (abs(weather - time_period) == z // 10) != (abs(weather - time_period) == z // 10 + 1):
        triggered.add(111)
    if (max(weather, time_period) == z // 10) != (max(weather, time_period) == z // 10 + 1):
        triggered.add(112)
    if (min(weather, time_period) * 10 <= z) != (min(weather, time_period) * 11 <= z):
        triggered.add(113)

    return triggered
def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """"""
    generated_path = execute_validation_rules(particle[0], particle[1], particle[2])

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # : x:1-50, y:1-50, z:1-50
        self.bounds = bounds if bounds else [(1, 100), (1, 100), (1, 100)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """target pathsPSO"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(particles[i][0], particles[i][1], particles[i][2]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(gbest_particle[0], gbest_particle[1], gbest_particle[2]),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """Path PSO"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - Path ")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations")
    print(f"Path : {len(target_paths)}")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"Path {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "" if result['success'] else f"({result['best_fitness']:.3f})"
        print(f"{status} | {result['time']:.2f}s | iterations{result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f": {success_count}/{len(target_paths)} ({success_rate:.1f}%) | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs}")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations, {len(target_paths)}Path ")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"---  {run_idx}/{num_runs} ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f": {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs} runcompleted | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 1: 
    ws1 = wb.active
    ws1.title = ""
    ws1.sheet_view.showGridLines = False

    headers = ["Run", "", "", "", "Average Iterations", "(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f" {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # 2: Path 
    ws2 = wb.create_sheet(title="Path ")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Path ID", "", "", "", "Average Iterations", "Minimum Iterations", "Maximum Iterations"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"Path  {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # 3: 
    ws3 = wb.create_sheet(title="")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Path ", "", "(x,y,z)", "", "Iterations", "Path "]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"{run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # 4: target paths
    ws4 = wb.create_sheet(title="target paths")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Path ID", "target paths", ""]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"Path  {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f" : {filename}")
    print(f"{'=' * 70}")
    print(f":")
    print(f"  1.        - {len(all_results)} run")
    print(f"  2. Path        - Path ")
    print(f"  3.    -  runPath ")
    print(f"  4. target paths       - target paths")
    print(f"{'=' * 70}\n")

    return filename


def main():


    target_paths = [
        [2, 3, 7, 12, 14, 17, 23, 24, 30, 31, 34, 35, 36, 38, 42, 45, 47, 75, 89, 92, 96, 97, 100, 101, 102, 107, 110,
         112],
        [3, 4, 8, 9, 14, 17, 24, 25, 30, 31, 32, 37, 68, 78, 82, 85, 89, 91, 93, 100, 101, 102, 106, 107, 109, 112],
        [2, 3, 7, 12, 15, 18, 23, 24, 30, 31, 34, 35, 36, 38, 42, 45, 47, 64, 75, 89, 92, 96, 97, 102, 103, 110],
        [2, 3, 9, 10, 15, 18, 23, 24, 30, 31, 33, 37, 46, 81, 86, 90, 91, 93, 98, 102, 103, 104, 106, 112],
        [2, 3, 7, 12, 20, 23, 24, 30, 31, 34, 35, 36, 38, 42, 45, 47, 89, 92, 96, 98, 102, 103, 106, 111],
        [2, 3, 11, 12, 14, 16, 23, 24, 30, 31, 34, 36, 38, 47, 86, 87, 90, 95, 100, 101, 103, 111, 113],
        [5, 6, 7, 12, 19, 26, 27, 29, 34, 35, 36, 38, 79, 89, 91, 94, 97, 101, 103, 105, 106, 111, 112],
        [1, 6, 11, 12, 15, 18, 27, 28, 29, 34, 36, 38, 41, 52, 84, 90, 91, 94, 98, 102, 103, 106, 111],
        [3, 4, 7, 12, 20, 24, 25, 30, 31, 34, 35, 36, 38, 45, 48, 50, 88, 89, 95, 97, 101, 103, 106],
        [5, 6, 7, 12, 13, 16, 22, 26, 27, 29, 34, 35, 36, 38, 79, 89, 91, 94, 98, 99, 101, 103, 106],
        [4, 5, 7, 12, 14, 17, 25, 26, 34, 35, 36, 38, 48, 77, 89, 97, 100, 101, 103, 106, 107, 111],
        [2, 3, 7, 12, 18, 23, 24, 30, 31, 34, 35, 36, 38, 42, 45, 47, 56, 75, 89, 92, 96, 101, 103],
        [2, 3, 11, 12, 13, 21, 23, 24, 30, 31, 34, 36, 38, 47, 87, 90, 95, 99, 102, 103, 106, 111],
        [1, 6, 9, 10, 13, 16, 22, 27, 28, 29, 33, 37, 40, 69, 90, 95, 98, 99, 101, 102, 106, 111],
        [1, 6, 7, 12, 15, 18, 27, 28, 29, 34, 35, 36, 38, 39, 41, 64, 79, 89, 97, 101, 103, 109],
        [1, 2, 11, 12, 15, 18, 23, 28, 34, 36, 38, 44, 54, 90, 93, 97, 101, 103, 104, 106, 111],
        [3, 4, 11, 12, 13, 21, 24, 25, 30, 31, 34, 36, 38, 50, 57, 90, 96, 99, 101, 103, 111],
        [1, 6, 8, 9, 14, 16, 27, 28, 29, 32, 37, 68, 76, 89, 92, 96, 100, 101, 102, 110, 112],
        [5, 6, 8, 9, 13, 16, 22, 26, 27, 29, 32, 37, 80, 89, 95, 98, 99, 102, 103, 106, 111],
        [1, 6, 10, 11, 14, 17, 27, 28, 29, 33, 40, 51, 84, 89, 98, 100, 101, 103, 106, 111],
        [3, 4, 10, 11, 13, 16, 22, 24, 25, 30, 31, 33, 49, 89, 97, 99, 102, 103, 106, 108],
        [1, 2, 10, 11, 13, 16, 22, 23, 28, 33, 43, 72, 89, 91, 95, 99, 101, 103, 111, 113],
        [3, 4, 11, 12, 20, 24, 25, 30, 31, 34, 36, 38, 50, 74, 90, 96, 97, 102, 103, 111],
        [4, 5, 8, 9, 15, 18, 25, 26, 32, 37, 60, 67, 82, 89, 96, 98, 101, 103, 106, 112],
        [2, 3, 10, 11, 18, 23, 24, 30, 31, 33, 46, 56, 71, 86, 89, 96, 101, 102, 112],
        [4, 5, 10, 11, 15, 18, 25, 26, 33, 83, 89, 97, 102, 103, 106, 109, 112],
        [4, 5, 11, 12, 14, 16, 25, 26, 34, 36, 38, 59, 90, 100, 102, 103, 106],
        [3, 4, 7, 8, 14, 17, 24, 25, 30, 31, 32, 35, 45, 48, 65, 77, 85, 88, 90, 96, 97, 100, 101, 103, 105, 106, 107,
         112],
        [3, 4, 7, 12, 15, 17, 24, 25, 30, 31, 34, 35, 36, 38, 45, 48, 50, 63, 77, 85, 89, 95, 101, 103, 110, 112],
        [2, 3, 8, 9, 13, 16, 22, 23, 24, 30, 31, 32, 37, 55, 78, 89, 97, 99, 102, 103, 105, 106, 109, 112],
        [1, 2, 9, 10, 13, 16, 22, 23, 28, 33, 37, 43, 69, 70, 90, 96, 97, 99, 102, 103, 106, 108, 111],
        [5, 6, 7, 12, 14, 17, 26, 27, 29, 34, 35, 36, 38, 62, 79, 89, 91, 94, 97, 100, 101, 102, 107],
        [1, 2, 9, 10, 14, 16, 23, 28, 33, 37, 43, 53, 90, 96, 100, 101, 103, 105, 106, 112],
        [3, 4, 7, 8, 18, 24, 25, 30, 31, 32, 35, 45, 48, 66, 77, 85, 90, 96, 102, 103, 110],
        [4, 5, 11, 12, 19, 25, 26, 34, 36, 38, 73, 83, 90, 98, 101, 102, 106, 112, 113],
        [3, 4, 10, 11, 15, 18, 24, 25, 30, 31, 33, 49, 58, 89, 97, 101, 102, 106, 112],
        [5, 6, 11, 12, 13, 21, 26, 27, 29, 34, 36, 38, 61, 90, 99, 102, 103, 109]
    ]

    print("=" * 70)
    print("baseline PSO")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("Program completed")


if __name__ == "__main__":
    main()