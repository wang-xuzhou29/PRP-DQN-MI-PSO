import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

def execute_validation_rules(x, y, weather):
    triggered = set()
    actions = []
    devices = {
        'main_light': 'green',
        'side_light': 'red',
        'pedestrian_light': 'red',
        'warning_system': 'off',
        'weather_alert': 'off'
    }

    # Fixed all if statements with proper syntax
    if (weather == 1 and x > 75 and y < 50) != (weather == 1 and x > 45 and y < 50):
        triggered.add(1)
    if (weather == 1 and x > 75 and y < 50) != (weather == 1 and x > 35 and y < 50):
        triggered.add(2)
    if (weather == 1 and x > 75 and y < 50) != (weather == 2 and x > 75 and y < 50):
        triggered.add(3)
    if (weather == 1 and x > 75 and y < 50) != (weather == 1 and x > 75 and y < 25):
        triggered.add(4)
    if (weather == 1 and x > 75 and y < 50) != (weather == 1 and x > 75 and y < 30):
        triggered.add(5)
    if (weather == 1 and x < 50 and y > 75) != (weather == 2 and x < 50 and y > 75):
        triggered.add(6)
    if (weather == 1 and x < 50 and y > 75) != (weather == 1 or x < 50 and y > 75):
        triggered.add(7)
    if (weather == 1 and x < 50 and y > 75) != (weather == 1 and x < 50 or y > 75):
        triggered.add(8)
    if (weather == 1 and x < 50 and y > 75) != (weather == 1 and x < 50 and y > 35):
        triggered.add(9)
    if (weather == 1 and x < 50 and y > 75) != (weather == 1 and x < 25 and y > 75):
        triggered.add(10)
    if (weather == 1 and x < 50 and y > 75) != (weather > 1 and x < 50 and y > 75):
        triggered.add(11)
    if (weather == 1 and x > 70 and y > 70) != (weather > 1 and x > 70 and y > 70):
        triggered.add(12)
    if (weather == 1 and x > 70 and y > 70) != (weather == 3 and x > 70 and y > 70):
        triggered.add(13)
    if (weather == 1 and x > 70 and y > 70) != (weather == 1 and x <= 70 and y > 70):
        triggered.add(14)
    if (weather == 1 and x > 70 and y > 70) != (weather == 1 and x > 70 and y <= 70):
        triggered.add(15)
    if (weather == 1 and x > 70 and y > 70) != (weather == 1 and x > 20 and y > 70):
        triggered.add(16)
    if (weather == 1 and x > 70 and y > 70) != (weather == 6 and x > 70 and y > 70):
        triggered.add(17)
    if (weather == 1 and x < 40 and y < 40) != (weather > 1 and x < 40 and y < 40):
        triggered.add(18)
    if (weather == 1 and x < 40 and y < 40) != (weather == 1 or x < 40 and y < 40):
        triggered.add(19)
    if (weather == 1 and x < 40 and y < 40) != (weather == 1 and x >= 40 and y < 40):
        triggered.add(20)
    if (weather == 1 and x < 40 and y < 40) != (weather == 1 and x < 20 and y < 40):
        triggered.add(21)
    if (weather == 1 and x < 40 and y < 40) != (weather == 1 and x < 40 or y < 40):
        triggered.add(22)
    if (weather == 1 and x < 40 and y < 40) != (weather == 1 and x < 40 and y >= 40):
        triggered.add(23)
    if (weather == 2 and x < 50 and y > 75) != (weather > 2 and x < 50 and y > 75):
        triggered.add(24)
    if (weather == 2 and x < 50 and y > 75) != (weather < 2 and x < 50 and y > 75):
        triggered.add(25)
    if (weather == 2 and x < 50 and y > 75) != (weather == 2 or x < 50 and y > 75):
        triggered.add(26)
    if (weather == 2 and x < 50 and y > 75) != (weather == 2 and x >= 50 and y > 75):
        triggered.add(27)
    if (weather == 2 and x < 50 and y > 75) != (weather == 2 and x < 50 or y > 75):
        triggered.add(28)
    if (weather == 2 and x > 85 and 45 < y < 70) != (weather > 2 and x > 85 and 45 < y < 70):
        triggered.add(29)
    if (weather == 2 and x > 85 and 45 < y < 70) != (weather < 2 and x > 85 and 45 < y < 70):
        triggered.add(30)
    if (weather == 2 and x > 85 and 45 < y < 70) != (weather == 2 or x > 85 and 45 < y < 70):
        triggered.add(31)
    if (weather == 2 and x > 85 and 45 < y < 70) != (weather == 2 and x > 85 or 45 < y < 70):
        triggered.add(32)
    if (weather == 2 and x > 85 and 45 < y < 70) != (weather == 2 and x > 85 and 60 < y < 70):
        triggered.add(33)
    if (weather == 2 and x > 85 and 45 < y < 70) != (weather == 2 and x > 85 and 45 < y < 80):
        triggered.add(34)
    if (weather == 3 and x > 75 and 40 < y < 65) != (weather > 3 and x > 75 and 40 < y < 65):
        triggered.add(35)
    if (weather == 3 and x > 75 and 40 < y < 65) != (weather < 3 and x > 75 and 40 < y < 65):
        triggered.add(36)
    if (weather == 3 and x > 75 and 40 < y < 65) != (weather == 3 or x > 75 and 40 < y < 65):
        triggered.add(37)
    if (weather == 3 and x > 75 and 40 < y < 65) != (weather == 3 and x > 75 and 50 < y < 65):
        triggered.add(38)
    if (weather == 3 and x > 75 and 40 < y < 65) != (weather == 3 and x > 75 and 40 < y < 75):
        triggered.add(39)
    if (weather == 3 and x > 75 and 40 < y < 65) != (weather == 3 and x > 45 and 40 < y < 65):
        triggered.add(40)
    if (weather == 4 and x > 65 and y > 65) != (weather > 4 and x > 65 and y > 65):
        triggered.add(41)
    if (weather == 4 and x > 65 and y > 65) != (weather < 4 and x > 65 and y > 65):
        triggered.add(42)
    if (weather == 4 and x > 65 and y > 65) != (weather == 4 or x > 65 and y > 65):
        triggered.add(43)
    if (weather == 4 and x > 65 and y > 65) != (weather == 4 and x > 25 and y > 65):
        triggered.add(44)
    if (weather == 4 and x > 65 and y > 65) != (weather == 4 and x > 65 and y > 35):
        triggered.add(45)
    if (weather == 4 and x > 65 and y > 65) != (weather == 4 and x > 65 and y > 85):
        triggered.add(46)
    if (weather == 5 and x < 45 and y > 75) != (weather > 5 and x < 45 and y > 75):
        triggered.add(47)
    if (weather == 5 and x < 45 and y > 75) != (weather < 5 and x < 45 and y > 75):
        triggered.add(48)
    if (weather == 5 and x < 45 and y > 75) != (weather == 6 and x < 45 and y > 75):
        triggered.add(49)
    if (weather == 5 and x < 45 and y > 75) != (weather == 5 or x < 45 and y > 75):
        triggered.add(50)
    if (weather == 5 and x < 45 and y > 75) != (weather == 5 and x < 25 and y > 75):
        triggered.add(51)
    if (weather == 5 and x < 45 and y > 75) != (weather == 5 and x < 35 and y > 75):
        triggered.add(52)
    if (weather == 5 and x < 45 and y > 75) != (weather == 5 and x < 45 and y > 65):
        triggered.add(53)
    if (weather == 5 and x < 45 and y > 75) != (weather == 5 and x < 45 and y > 55):
        triggered.add(54)
    if (weather == 5 and x < 45 and y > 75) != (weather == 5 and x < 45 and y > 35):
        triggered.add(55)
    if (weather == 5 and x < 45 and y > 75) != (weather == 5 and x < 45 and y > 25):
        triggered.add(56)
    if (weather == 5 and x < 45 and y > 75) != (weather == 5 and x < 45 and y > 15):
        triggered.add(57)
    if (weather == 6 and x < 40 and y < 40) != (weather < 6 and x < 40 and y < 40):
        triggered.add(58)
    if (weather == 6 and x < 40 and y < 40) != (weather != 6 and x < 40 and y < 40):
        triggered.add(59)
    if (weather == 6 and x < 40 and y < 40) != (weather == 6 and x < 20 and y < 40):
        triggered.add(60)
    if (weather == 6 and x < 40 and y < 40) != (weather == 6 and x < 40 and y < 20):
        triggered.add(61)
    if (weather == 6 and x < 40 and y < 40) != (weather == 6 and x < 40 or y < 40):
        triggered.add(62)
    if (weather == 1 and x > 90) != (weather > 1 and x > 90):
        triggered.add(63)
    if (weather == 1 and x > 90) != (weather == 1 and y > 90):
        triggered.add(64)
    if (weather == 1 and x > 90) != (weather == 1 and x > 40):
        triggered.add(65)
    if (weather == 1 and x > 90) != (weather == 1 and x > 20):
        triggered.add(66)
    if (weather == 1 and x > 90) != (weather == 1 and x > 60):
        triggered.add(67)
    if (weather in [2, 3, 4, 6] and x > 80) != (weather in [1, 3, 4, 6] and x > 80):
        triggered.add(68)
    if (weather in [2, 3, 4, 6] and x > 80) != (weather in [3, 4, 6] and x > 80):
        triggered.add(69)
    if (weather in [2, 3, 4, 6] and x > 80) != (weather in [2, 4, 6] and x > 80):
        triggered.add(70)
    if (weather in [2, 3, 4, 6] and x > 80) != (weather in [2, 3, 6] and x > 80):
        triggered.add(71)
    if (weather in [2, 3, 4, 6] and x > 80) != (weather in [2, 3, 4] and x > 80):
        triggered.add(72)
    if (weather in [2, 3, 4, 6] and x > 80) != (weather in [1, 2, 3, 4, 6] and x > 80):
        triggered.add(73)
    if (weather in [2, 3, 4, 6] and x > 80) != (weather in [2, 1, 4, 6] and x > 80):
        triggered.add(74)
    if (weather in [2, 3, 4, 6] and x > 80) != (weather in [2, 3, 1, 6] and x > 80):
        triggered.add(75)
    if (weather in [2, 3, 4, 6] and x > 80) != (weather in [2, 3, 4, 1] and x > 80):
        triggered.add(76)
    if (weather in [2, 3, 4, 6] and x > 80) != (weather in [5, 3, 4, 6] and x > 80):
        triggered.add(77)
    if (weather in [2, 3, 4, 6] and x > 80) != (weather in [2, 3, 4, 6] and x > 60):
        triggered.add(78)
    if (weather in [2, 3, 4, 6] and x > 80) != (weather in [2, 3, 4, 6] and x > 30):
        triggered.add(79)
    if (weather in [3, 4] and 60 < x < 85) != (weather in [4] and 60 < x < 85):
        triggered.add(80)
    if (weather in [3, 4] and 60 < x < 85) != (weather in [3] and 60 < x < 85):
        triggered.add(81)
    if (weather in [3, 4] and 60 < x < 85) != (weather in [3, 4] or 60 < x < 85):
        triggered.add(82)
    if (weather in [3, 4] and 60 < x < 85) != (weather in [3, 4] and 50 < x < 85):
        triggered.add(83)
    if (weather in [3, 4] and 60 < x < 85) != (weather in [3, 4] and 20 < x < 85):
        triggered.add(84)
    if (weather in [3, 4] and 60 < x < 85) != (weather in [3, 4] and 60 < x < 75):
        triggered.add(85)
    if (weather in [3, 4] and 60 < x < 85) != (weather in [3, 4] and 60 < x < 65):
        triggered.add(86)
    if (weather in [3, 4] and 60 < x < 85) != (weather in [3, 4, 5] and 60 < x < 85):
        triggered.add(87)
    if (weather in [3, 4] and 60 < x < 85) != (weather in [3, 4, 2] and 60 < x < 85):
        triggered.add(88)
    if (weather == 2 and 45 < x < 70) != (weather > 2 and 45 < x < 70):
        triggered.add(89)
    if (weather == 2 and 45 < x < 70) != (weather < 2 and 45 < x < 70):
        triggered.add(90)
    if (weather == 2 and 45 < x < 70) != (weather == 3 and 45 < x < 70):
        triggered.add(91)
    if (weather == 2 and 45 < x < 70) != (weather == 5 and 45 < x < 70):
        triggered.add(92)
    if (weather == 2 and 45 < x < 70) != (weather == 2 and 55 < x < 70):
        triggered.add(93)
    if (weather == 2 and 45 < x < 70) != (weather == 2 and 45 < y < 70):
        triggered.add(94)
    if (weather == 2 and 45 < x < 70) != (weather == 2 and 45 < x < 60):
        triggered.add(95)
    if (weather == 2 and 45 < x < 70) != (weather == 2 and 45 < x < 50):
        triggered.add(96)
    if (x - y > 60 and x > 70) != (x + y > 60 and x > 70):
        triggered.add(97)
    if (x - y > 60 and x > 70) != (x - y > 60 and y > 70):
        triggered.add(98)
    if (x - y > 60 and x > 70) != (x - y > 60 or x > 70):
        triggered.add(99)
    if (x - y > 60 and x > 70) != (x - y > 50 and x > 70):
        triggered.add(100)
    if (x - y > 60 and x > 70) != (x - y > 30 and x > 70):
        triggered.add(101)
    if (x - y > 60 and x > 70) != (x - y > 20 and x > 70):
        triggered.add(102)
    if (x - y > 60 and x > 70) != (x - y > 60 and x < 70):
        triggered.add(103)
    if (x - y > 60 and x > 70) != (x - y < 60 and x > 70):
        triggered.add(104)
    if (x > 90 and y > 90 and abs(x - y) < 10) != (x > 50 and y > 90 and abs(x - y) < 10):
        triggered.add(105)
    if (x > 90 and y > 90 and abs(x - y) < 10) != (x > 30 and y > 90 and abs(x - y) < 10):
        triggered.add(106)
    if (x > 90 and y > 90 and abs(x - y) < 10) != (x < 90 and y > 90 and abs(x - y) < 10):
        triggered.add(107)
    if (x > 90 and y > 90 and abs(x - y) < 10) != (x > 90 and y < 90 and abs(x - y) < 10):
        triggered.add(108)
    if (x > 90 and y > 90 and abs(x - y) < 10) != (x > 90 and y > 30 and abs(x - y) < 10):
        triggered.add(109)
    if (x > 90 and y > 90 and abs(x - y) < 10) != (x > 90 and y > 90 and abs(x + y) < 10):
        triggered.add(110)
    if (x > 90 and y > 90 and abs(x - y) < 10) != (x > 90 and y > 90 and abs(x - y) > 10):
        triggered.add(111)
    if (25 < x < 45 and 25 < y < 45 and abs(x - y) < 12) != (15 < x < 45 and 25 < y < 45 and abs(x - y) < 12):
        triggered.add(112)
    if (25 < x < 45 and 25 < y < 45 and abs(x - y) < 12) != (25 < x < 45 or 25 < y < 45 and abs(x - y) < 12):
        triggered.add(113)
    if (25 < x < 45 and 25 < y < 45 and abs(x - y) < 12) != (25 < x - weather < 45 and 25 < y < 45 and abs(x - y) < 12):
        triggered.add(114)
    if (25 < x < 45 and 25 < y < 45 and abs(x - y) < 12) != (25 < x < 45 and 25 < y - weather < 45 and abs(x - y) < 12):
        triggered.add(115)
    if (25 < x < 45 and 25 < y < 45 and abs(x - y) < 12) != (25 < x + weather < 45 and 25 < y < 45 and abs(x - y) < 12):
        triggered.add(116)
    if (25 < x < 45 and 25 < y < 45 and abs(x - y) < 12) != (25 < x < 45 and 25 < y + weather < 45 and abs(x - y) < 12):
        triggered.add(117)
    if (25 < x < 45 and 25 < y < 45 and abs(x - y) < 12) != (25 < x < 45 and 25 < y < 45 or abs(x - y) < 12):
        triggered.add(118)
    if (25 < x < 45 and 25 < y < 45 and abs(x - y) < 12) != (25 < x < 45 and 25 < y < 45 and abs(x - y) < 22):
        triggered.add(119)
    if (25 < x < 45 and 25 < y < 45 and abs(x - y) < 12) != (25 < x < 45 and 25 < y < 45 and abs(x - weather) < 12):
        triggered.add(120)

    return triggered, actions, devices


def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """"""
    generated_path = execute_validation_rules(particle[0], particle[1], particle[2])

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # : x:1-50, y:1-50, z:1-50
        self.bounds = bounds if bounds else [(1, 100), (1, 100), (1, 6)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """target pathsPSO"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(particles[i][0], particles[i][1], particles[i][2]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(gbest_particle[0], gbest_particle[1], gbest_particle[2]),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """Path PSO"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - Path ")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations")
    print(f"Path : {len(target_paths)}")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"Path {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "" if result['success'] else f"({result['best_fitness']:.3f})"
        print(f"{status} | {result['time']:.2f}s | iterations{result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f": {success_count}/{len(target_paths)} ({success_rate:.1f}%) | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs}")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations, {len(target_paths)}Path ")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"---  {run_idx}/{num_runs} ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f": {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs} runcompleted | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 1:
    ws1 = wb.active
    ws1.title = ""
    ws1.sheet_view.showGridLines = False

    headers = ["Run", "", "", "", "Average Iterations", "(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f" {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # 2: Path
    ws2 = wb.create_sheet(title="Path ")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Path ID", "", "", "", "Average Iterations", "Minimum Iterations", "Maximum Iterations"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"Path  {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # 3:
    ws3 = wb.create_sheet(title="")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Path ", "", "(x,y,z)", "", "Iterations", "Path "]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"{run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # 4: target paths
    ws4 = wb.create_sheet(title="target paths")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Path ID", "target paths", ""]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"Path  {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f" : {filename}")
    print(f"{'=' * 70}")
    print(f":")
    print(f"  1.        - {len(all_results)} run")
    print(f"  2. Path        - Path ")
    print(f"  3.    -  runPath ")
    print(f"  4. target paths       - target paths")
    print(f"{'=' * 70}\n")

    return filename


def main():


    target_paths = [
        [7, 8, 12, 13, 14, 15, 17, 19, 28, 42, 43, 64, 65, 66, 67, 68, 73, 74, 75, 76, 97, 99, 104, 105, 106, 107, 118],
        [3, 4, 5, 7, 15, 19, 20, 22, 62, 65, 66, 67, 68, 73, 74, 75, 76, 82, 97, 99, 100, 101, 102, 104],
        [3, 4, 5, 7, 15, 19, 30, 31, 32, 36, 37, 63, 64, 68, 73, 74, 75, 76, 97, 99, 100, 101, 102, 104],
        [7, 8, 12, 13, 14, 15, 17, 19, 28, 42, 43, 63, 64, 68, 73, 74, 75, 76, 97, 99, 104, 109, 118],
        [8, 12, 13, 28, 37, 42, 43, 70, 74, 80, 85, 86, 97, 99, 104, 105, 106, 107, 118],
        [8, 12, 26, 27, 28, 31, 32, 34, 42, 43, 63, 68, 69, 77, 97, 99, 102, 104],
        [8, 12, 28, 41, 42, 71, 75, 81, 85, 86, 97, 99, 104, 105, 106, 107, 118],
        [8, 26, 27, 28, 31, 78, 79, 82, 88, 89, 90, 91, 92, 94, 95, 96, 118],
        [29, 31, 32, 41, 42, 46, 63, 71, 75, 82, 97, 99, 101, 102, 104],
        [6, 10, 11, 14, 16, 19, 25, 26, 28, 48, 50, 64, 65, 66, 113],
        [29, 31, 32, 35, 37, 63, 72, 76, 97, 99, 100, 101, 102, 104],
        [29, 31, 32, 37, 39, 63, 70, 74, 82, 97, 99, 101, 102, 104],
        [32, 35, 36, 38, 70, 74, 80, 85, 86, 97, 99, 101, 102, 104],
        [26, 29, 30, 33, 36, 37, 68, 69, 77, 94, 97, 99, 102, 104],
        [7, 8, 9, 18, 20, 21, 23, 58, 59, 62, 66, 113, 118, 119],
        [1, 2, 7, 15, 19, 65, 66, 67, 82, 97, 99, 101, 102, 104],
        [18, 19, 22, 26, 28, 31, 58, 59, 62, 112, 113, 116, 118],
        [18, 19, 22, 37, 58, 59, 62, 79, 82, 84, 113, 118, 119],
        [22, 43, 45, 62, 63, 71, 75, 82, 97, 99, 100, 101, 102],
        [18, 19, 22, 37, 58, 59, 62, 79, 82, 84, 113, 117, 118],
        [26, 28, 31, 79, 89, 90, 91, 92, 93, 94, 113, 114, 118],
        [18, 19, 22, 26, 28, 31, 58, 59, 62, 114, 115, 120],
        [37, 40, 78, 79, 80, 85, 86, 97, 99, 101, 102, 104],
        [18, 19, 22, 50, 56, 57, 58, 59, 62, 112, 113, 118],
        [22, 50, 62, 77, 82, 87, 97, 99, 100, 101, 102],
        [18, 19, 22, 58, 59, 60, 61, 79, 115, 116, 120],
        [22, 43, 62, 78, 79, 81, 89, 99, 103],
        [37, 40, 79, 82, 83, 84, 89, 91, 118],
        [22, 50, 55, 56, 57, 62, 116, 120],
        [43, 44, 78, 79, 81, 86, 89, 118],
        [7, 8, 12, 13, 14, 15, 17, 19, 28, 42, 43, 63, 64, 68, 73, 74, 75, 76, 97, 99, 104, 108, 109, 118],
        [7, 8, 11, 24, 26, 28, 37, 79, 82, 84],
        [22, 43, 45, 62, 63, 71, 75, 82, 98, 103, 104],
        [7, 8, 11, 24, 26, 28, 47, 48, 49, 51],
        [8, 12, 28, 41, 43, 50, 63, 77, 97, 99, 104, 107, 108, 110, 111, 118],
        [32, 50, 53, 54, 55, 56, 57, 113],
        [7, 8, 11, 24, 26, 28, 47, 48, 49, 51, 52, 113]
    ]

    print("=" * 70)
    print("baseline PSO")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("Program completed")


if __name__ == "__main__":
    main()