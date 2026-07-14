import os
import random
from collections import deque
import numpy as np
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"


def generate_input():
    return [
        random.randint(1, 100),
        random.randint(1, 100),
        random.randint(1, 6)
    ]


def execute_validation_rules(x, y, weather):
    triggered = set()
    actions = []
    devices = {
        'main_light': 'green',
        'side_light': 'red',
        'pedestrian_light': 'red',
        'warning_system': 'off',
        'weather_alert': 'off'
    }

    # Fixed all if statements with proper syntax
    if (weather == 1 and x > 75 and y < 50) != (weather == 1 and x > 45 and y < 50):
        triggered.add(1)
    if (weather == 1 and x > 75 and y < 50) != (weather == 1 and x > 35 and y < 50):
        triggered.add(2)
    if (weather == 1 and x > 75 and y < 50) != (weather == 2 and x > 75 and y < 50):
        triggered.add(3)
    if (weather == 1 and x > 75 and y < 50) != (weather == 1 and x > 75 and y < 25):
        triggered.add(4)
    if (weather == 1 and x > 75 and y < 50) != (weather == 1 and x > 75 and y < 30):
        triggered.add(5)
    if (weather == 1 and x < 50 and y > 75) != (weather == 2 and x < 50 and y > 75):
        triggered.add(6)
    if (weather == 1 and x < 50 and y > 75) != (weather == 1 or x < 50 and y > 75):
        triggered.add(7)
    if (weather == 1 and x < 50 and y > 75) != (weather == 1 and x < 50 or y > 75):
        triggered.add(8)
    if (weather == 1 and x < 50 and y > 75) != (weather == 1 and x < 50 and y > 35):
        triggered.add(9)
    if (weather == 1 and x < 50 and y > 75) != (weather == 1 and x < 25 and y > 75):
        triggered.add(10)
    if (weather == 1 and x < 50 and y > 75) != (weather > 1 and x < 50 and y > 75):
        triggered.add(11)
    if (weather == 1 and x > 70 and y > 70) != (weather > 1 and x > 70 and y > 70):
        triggered.add(12)
    if (weather == 1 and x > 70 and y > 70) != (weather == 3 and x > 70 and y > 70):
        triggered.add(13)
    if (weather == 1 and x > 70 and y > 70) != (weather == 1 and x <= 70 and y > 70):
        triggered.add(14)
    if (weather == 1 and x > 70 and y > 70) != (weather == 1 and x > 70 and y <= 70):
        triggered.add(15)
    if (weather == 1 and x > 70 and y > 70) != (weather == 1 and x > 20 and y > 70):
        triggered.add(16)
    if (weather == 1 and x > 70 and y > 70) != (weather == 6 and x > 70 and y > 70):
        triggered.add(17)
    if (weather == 1 and x < 40 and y < 40) != (weather > 1 and x < 40 and y < 40):
        triggered.add(18)
    if (weather == 1 and x < 40 and y < 40) != (weather == 1 or x < 40 and y < 40):
        triggered.add(19)
    if (weather == 1 and x < 40 and y < 40) != (weather == 1 and x >= 40 and y < 40):
        triggered.add(20)
    if (weather == 1 and x < 40 and y < 40) != (weather == 1 and x < 20 and y < 40):
        triggered.add(21)
    if (weather == 1 and x < 40 and y < 40) != (weather == 1 and x < 40 or y < 40):
        triggered.add(22)
    if (weather == 1 and x < 40 and y < 40) != (weather == 1 and x < 40 and y >= 40):
        triggered.add(23)
    if (weather == 2 and x < 50 and y > 75) != (weather > 2 and x < 50 and y > 75):
        triggered.add(24)
    if (weather == 2 and x < 50 and y > 75) != (weather < 2 and x < 50 and y > 75):
        triggered.add(25)
    if (weather == 2 and x < 50 and y > 75) != (weather == 2 or x < 50 and y > 75):
        triggered.add(26)
    if (weather == 2 and x < 50 and y > 75) != (weather == 2 and x >= 50 and y > 75):
        triggered.add(27)
    if (weather == 2 and x < 50 and y > 75) != (weather == 2 and x < 50 or y > 75):
        triggered.add(28)
    if (weather == 2 and x > 85 and 45 < y < 70) != (weather > 2 and x > 85 and 45 < y < 70):
        triggered.add(29)
    if (weather == 2 and x > 85 and 45 < y < 70) != (weather < 2 and x > 85 and 45 < y < 70):
        triggered.add(30)
    if (weather == 2 and x > 85 and 45 < y < 70) != (weather == 2 or x > 85 and 45 < y < 70):
        triggered.add(31)
    if (weather == 2 and x > 85 and 45 < y < 70) != (weather == 2 and x > 85 or 45 < y < 70):
        triggered.add(32)
    if (weather == 2 and x > 85 and 45 < y < 70) != (weather == 2 and x > 85 and 60 < y < 70):
        triggered.add(33)
    if (weather == 2 and x > 85 and 45 < y < 70) != (weather == 2 and x > 85 and 45 < y < 80):
        triggered.add(34)
    if (weather == 3 and x > 75 and 40 < y < 65) != (weather > 3 and x > 75 and 40 < y < 65):
        triggered.add(35)
    if (weather == 3 and x > 75 and 40 < y < 65) != (weather < 3 and x > 75 and 40 < y < 65):
        triggered.add(36)
    if (weather == 3 and x > 75 and 40 < y < 65) != (weather == 3 or x > 75 and 40 < y < 65):
        triggered.add(37)
    if (weather == 3 and x > 75 and 40 < y < 65) != (weather == 3 and x > 75 and 50 < y < 65):
        triggered.add(38)
    if (weather == 3 and x > 75 and 40 < y < 65) != (weather == 3 and x > 75 and 40 < y < 75):
        triggered.add(39)
    if (weather == 3 and x > 75 and 40 < y < 65) != (weather == 3 and x > 45 and 40 < y < 65):
        triggered.add(40)
    if (weather == 4 and x > 65 and y > 65) != (weather > 4 and x > 65 and y > 65):
        triggered.add(41)
    if (weather == 4 and x > 65 and y > 65) != (weather < 4 and x > 65 and y > 65):
        triggered.add(42)
    if (weather == 4 and x > 65 and y > 65) != (weather == 4 or x > 65 and y > 65):
        triggered.add(43)
    if (weather == 4 and x > 65 and y > 65) != (weather == 4 and x > 25 and y > 65):
        triggered.add(44)
    if (weather == 4 and x > 65 and y > 65) != (weather == 4 and x > 65 and y > 35):
        triggered.add(45)
    if (weather == 4 and x > 65 and y > 65) != (weather == 4 and x > 65 and y > 85):
        triggered.add(46)
    if (weather == 5 and x < 45 and y > 75) != (weather > 5 and x < 45 and y > 75):
        triggered.add(47)
    if (weather == 5 and x < 45 and y > 75) != (weather < 5 and x < 45 and y > 75):
        triggered.add(48)
    if (weather == 5 and x < 45 and y > 75) != (weather == 6 and x < 45 and y > 75):
        triggered.add(49)
    if (weather == 5 and x < 45 and y > 75) != (weather == 5 or x < 45 and y > 75):
        triggered.add(50)
    if (weather == 5 and x < 45 and y > 75) != (weather == 5 and x < 25 and y > 75):
        triggered.add(51)
    if (weather == 5 and x < 45 and y > 75) != (weather == 5 and x < 35 and y > 75):
        triggered.add(52)
    if (weather == 5 and x < 45 and y > 75) != (weather == 5 and x < 45 and y > 65):
        triggered.add(53)
    if (weather == 5 and x < 45 and y > 75) != (weather == 5 and x < 45 and y > 55):
        triggered.add(54)
    if (weather == 5 and x < 45 and y > 75) != (weather == 5 and x < 45 and y > 35):
        triggered.add(55)
    if (weather == 5 and x < 45 and y > 75) != (weather == 5 and x < 45 and y > 25):
        triggered.add(56)
    if (weather == 5 and x < 45 and y > 75) != (weather == 5 and x < 45 and y > 15):
        triggered.add(57)
    if (weather == 6 and x < 40 and y < 40) != (weather < 6 and x < 40 and y < 40):
        triggered.add(58)
    if (weather == 6 and x < 40 and y < 40) != (weather != 6 and x < 40 and y < 40):
        triggered.add(59)
    if (weather == 6 and x < 40 and y < 40) != (weather == 6 and x < 20 and y < 40):
        triggered.add(60)
    if (weather == 6 and x < 40 and y < 40) != (weather == 6 and x < 40 and y < 20):
        triggered.add(61)
    if (weather == 6 and x < 40 and y < 40) != (weather == 6 and x < 40 or y < 40):
        triggered.add(62)
    if (weather == 1 and x > 90) != (weather > 1 and x > 90):
        triggered.add(63)
    if (weather == 1 and x > 90) != (weather == 1 and y > 90):
        triggered.add(64)
    if (weather == 1 and x > 90) != (weather == 1 and x > 40):
        triggered.add(65)
    if (weather == 1 and x > 90) != (weather == 1 and x > 20):
        triggered.add(66)
    if (weather == 1 and x > 90) != (weather == 1 and x > 60):
        triggered.add(67)
    if (weather in [2, 3, 4, 6] and x > 80) != (weather in [1, 3, 4, 6] and x > 80):
        triggered.add(68)
    if (weather in [2, 3, 4, 6] and x > 80) != (weather in [3, 4, 6] and x > 80):
        triggered.add(69)
    if (weather in [2, 3, 4, 6] and x > 80) != (weather in [2, 4, 6] and x > 80):
        triggered.add(70)
    if (weather in [2, 3, 4, 6] and x > 80) != (weather in [2, 3, 6] and x > 80):
        triggered.add(71)
    if (weather in [2, 3, 4, 6] and x > 80) != (weather in [2, 3, 4] and x > 80):
        triggered.add(72)
    if (weather in [2, 3, 4, 6] and x > 80) != (weather in [1, 2, 3, 4, 6] and x > 80):
        triggered.add(73)
    if (weather in [2, 3, 4, 6] and x > 80) != (weather in [2, 1, 4, 6] and x > 80):
        triggered.add(74)
    if (weather in [2, 3, 4, 6] and x > 80) != (weather in [2, 3, 1, 6] and x > 80):
        triggered.add(75)
    if (weather in [2, 3, 4, 6] and x > 80) != (weather in [2, 3, 4, 1] and x > 80):
        triggered.add(76)
    if (weather in [2, 3, 4, 6] and x > 80) != (weather in [5, 3, 4, 6] and x > 80):
        triggered.add(77)
    if (weather in [2, 3, 4, 6] and x > 80) != (weather in [2, 3, 4, 6] and x > 60):
        triggered.add(78)
    if (weather in [2, 3, 4, 6] and x > 80) != (weather in [2, 3, 4, 6] and x > 30):
        triggered.add(79)
    if (weather in [3, 4] and 60 < x < 85) != (weather in [4] and 60 < x < 85):
        triggered.add(80)
    if (weather in [3, 4] and 60 < x < 85) != (weather in [3] and 60 < x < 85):
        triggered.add(81)
    if (weather in [3, 4] and 60 < x < 85) != (weather in [3, 4] or 60 < x < 85):
        triggered.add(82)
    if (weather in [3, 4] and 60 < x < 85) != (weather in [3, 4] and 50 < x < 85):
        triggered.add(83)
    if (weather in [3, 4] and 60 < x < 85) != (weather in [3, 4] and 20 < x < 85):
        triggered.add(84)
    if (weather in [3, 4] and 60 < x < 85) != (weather in [3, 4] and 60 < x < 75):
        triggered.add(85)
    if (weather in [3, 4] and 60 < x < 85) != (weather in [3, 4] and 60 < x < 65):
        triggered.add(86)
    if (weather in [3, 4] and 60 < x < 85) != (weather in [3, 4, 5] and 60 < x < 85):
        triggered.add(87)
    if (weather in [3, 4] and 60 < x < 85) != (weather in [3, 4, 2] and 60 < x < 85):
        triggered.add(88)
    if (weather == 2 and 45 < x < 70) != (weather > 2 and 45 < x < 70):
        triggered.add(89)
    if (weather == 2 and 45 < x < 70) != (weather < 2 and 45 < x < 70):
        triggered.add(90)
    if (weather == 2 and 45 < x < 70) != (weather == 3 and 45 < x < 70):
        triggered.add(91)
    if (weather == 2 and 45 < x < 70) != (weather == 5 and 45 < x < 70):
        triggered.add(92)
    if (weather == 2 and 45 < x < 70) != (weather == 2 and 55 < x < 70):
        triggered.add(93)
    if (weather == 2 and 45 < x < 70) != (weather == 2 and 45 < y < 70):
        triggered.add(94)
    if (weather == 2 and 45 < x < 70) != (weather == 2 and 45 < x < 60):
        triggered.add(95)
    if (weather == 2 and 45 < x < 70) != (weather == 2 and 45 < x < 50):
        triggered.add(96)
    if (x - y > 60 and x > 70) != (x + y > 60 and x > 70):
        triggered.add(97)
    if (x - y > 60 and x > 70) != (x - y > 60 and y > 70):
        triggered.add(98)
    if (x - y > 60 and x > 70) != (x - y > 60 or x > 70):
        triggered.add(99)
    if (x - y > 60 and x > 70) != (x - y > 50 and x > 70):
        triggered.add(100)
    if (x - y > 60 and x > 70) != (x - y > 30 and x > 70):
        triggered.add(101)
    if (x - y > 60 and x > 70) != (x - y > 20 and x > 70):
        triggered.add(102)
    if (x - y > 60 and x > 70) != (x - y > 60 and x < 70):
        triggered.add(103)
    if (x - y > 60 and x > 70) != (x - y < 60 and x > 70):
        triggered.add(104)
    if (x > 90 and y > 90 and abs(x - y) < 10) != (x > 50 and y > 90 and abs(x - y) < 10):
        triggered.add(105)
    if (x > 90 and y > 90 and abs(x - y) < 10) != (x > 30 and y > 90 and abs(x - y) < 10):
        triggered.add(106)
    if (x > 90 and y > 90 and abs(x - y) < 10) != (x < 90 and y > 90 and abs(x - y) < 10):
        triggered.add(107)
    if (x > 90 and y > 90 and abs(x - y) < 10) != (x > 90 and y < 90 and abs(x - y) < 10):
        triggered.add(108)
    if (x > 90 and y > 90 and abs(x - y) < 10) != (x > 90 and y > 30 and abs(x - y) < 10):
        triggered.add(109)
    if (x > 90 and y > 90 and abs(x - y) < 10) != (x > 90 and y > 90 and abs(x + y) < 10):
        triggered.add(110)
    if (x > 90 and y > 90 and abs(x - y) < 10) != (x > 90 and y > 90 and abs(x - y) > 10):
        triggered.add(111)
    if (25 < x < 45 and 25 < y < 45 and abs(x - y) < 12) != (15 < x < 45 and 25 < y < 45 and abs(x - y) < 12):
        triggered.add(112)
    if (25 < x < 45 and 25 < y < 45 and abs(x - y) < 12) != (25 < x < 45 or 25 < y < 45 and abs(x - y) < 12):
        triggered.add(113)
    if (25 < x < 45 and 25 < y < 45 and abs(x - y) < 12) != (25 < x - weather < 45 and 25 < y < 45 and abs(x - y) < 12):
        triggered.add(114)
    if (25 < x < 45 and 25 < y < 45 and abs(x - y) < 12) != (25 < x < 45 and 25 < y - weather < 45 and abs(x - y) < 12):
        triggered.add(115)
    if (25 < x < 45 and 25 < y < 45 and abs(x - y) < 12) != (25 < x + weather < 45 and 25 < y < 45 and abs(x - y) < 12):
        triggered.add(116)
    if (25 < x < 45 and 25 < y < 45 and abs(x - y) < 12) != (25 < x < 45 and 25 < y + weather < 45 and abs(x - y) < 12):
        triggered.add(117)
    if (25 < x < 45 and 25 < y < 45 and abs(x - y) < 12) != (25 < x < 45 and 25 < y < 45 or abs(x - y) < 12):
        triggered.add(118)
    if (25 < x < 45 and 25 < y < 45 and abs(x - y) < 12) != (25 < x < 45 and 25 < y < 45 and abs(x - y) < 22):
        triggered.add(119)
    if (25 < x < 45 and 25 < y < 45 and abs(x - y) < 12) != (25 < x < 45 and 25 < y < 45 and abs(x - weather) < 12):
        triggered.add(120)

    return triggered, actions, devices


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0


# Fixed: Changed to proper list syntax (removed A1 =, etc.)
TARGET_PATHS = [
    [7, 8, 12, 13, 14, 15, 17, 19, 28, 42, 43, 64, 65, 66, 67, 68, 73, 74, 75, 76, 97, 99, 104, 105, 106, 107, 118],
    [3, 4, 5, 7, 15, 19, 20, 22, 62, 65, 66, 67, 68, 73, 74, 75, 76, 82, 97, 99, 100, 101, 102, 104],
    [3, 4, 5, 7, 15, 19, 30, 31, 32, 36, 37, 63, 64, 68, 73, 74, 75, 76, 97, 99, 100, 101, 102, 104],
    [7, 8, 12, 13, 14, 15, 17, 19, 28, 42, 43, 63, 64, 68, 73, 74, 75, 76, 97, 99, 104, 109, 118],
    [8, 12, 13, 28, 37, 42, 43, 70, 74, 80, 85, 86, 97, 99, 104, 105, 106, 107, 118],
    [8, 12, 26, 27, 28, 31, 32, 34, 42, 43, 63, 68, 69, 77, 97, 99, 102, 104],
    [8, 12, 28, 41, 42, 71, 75, 81, 85, 86, 97, 99, 104, 105, 106, 107, 118],
    [8, 26, 27, 28, 31, 78, 79, 82, 88, 89, 90, 91, 92, 94, 95, 96, 118],
    [29, 31, 32, 41, 42, 46, 63, 71, 75, 82, 97, 99, 101, 102, 104],
    [6, 10, 11, 14, 16, 19, 25, 26, 28, 48, 50, 64, 65, 66, 113],
    [29, 31, 32, 35, 37, 63, 72, 76, 97, 99, 100, 101, 102, 104],
    [29, 31, 32, 37, 39, 63, 70, 74, 82, 97, 99, 101, 102, 104],
    [32, 35, 36, 38, 70, 74, 80, 85, 86, 97, 99, 101, 102, 104],
    [26, 29, 30, 33, 36, 37, 68, 69, 77, 94, 97, 99, 102, 104],
    [7, 8, 9, 18, 20, 21, 23, 58, 59, 62, 66, 113, 118, 119],
    [1, 2, 7, 15, 19, 65, 66, 67, 82, 97, 99, 101, 102, 104],
    [18, 19, 22, 26, 28, 31, 58, 59, 62, 112, 113, 116, 118],
    [18, 19, 22, 37, 58, 59, 62, 79, 82, 84, 113, 118, 119],
    [22, 43, 45, 62, 63, 71, 75, 82, 97, 99, 100, 101, 102],
    [18, 19, 22, 37, 58, 59, 62, 79, 82, 84, 113, 117, 118],
    [26, 28, 31, 79, 89, 90, 91, 92, 93, 94, 113, 114, 118],
    [18, 19, 22, 26, 28, 31, 58, 59, 62, 114, 115, 120],
    [37, 40, 78, 79, 80, 85, 86, 97, 99, 101, 102, 104],
    [18, 19, 22, 50, 56, 57, 58, 59, 62, 112, 113, 118],
    [22, 50, 62, 77, 82, 87, 97, 99, 100, 101, 102],
    [18, 19, 22, 58, 59, 60, 61, 79, 115, 116, 120],
    [22, 43, 62, 78, 79, 81, 89, 99, 103],
    [37, 40, 79, 82, 83, 84, 89, 91, 118],
    [22, 50, 55, 56, 57, 62, 116, 120],
    [43, 44, 78, 79, 81, 86, 89, 118],
    [7, 8, 12, 13, 14, 15, 17, 19, 28, 42, 43, 63, 64, 68, 73, 74, 75, 76, 97, 99, 104, 108, 109, 118],
    [7, 8, 11, 24, 26, 28, 37, 79, 82, 84],
    [22, 43, 45, 62, 63, 71, 75, 82, 98, 103, 104],
    [7, 8, 11, 24, 26, 28, 47, 48, 49, 51],
    [8, 12, 28, 41, 43, 50, 63, 77, 97, 99, 104, 107, 108, 110, 111, 118],
    [32, 50, 53, 54, 55, 56, 57, 113],
    [7, 8, 11, 24, 26, 28, 47, 48, 49, 51, 52, 113]
]


class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 2000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(TARGET_PATHS)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }


def compute_robustness(state, path):
    # Fixed: Changed execute_Tr to execute_validation_rules and properly handle tuple return
    base, actions, devices = execute_validation_rules(state[0], state[1], state[2])
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                    np.clip(state[0] + dx, 1, 100),  # x range 1-100
                    np.clip(state[1] + dy, 1, 100),  # y range 1-100
                    np.clip(state[2] + dz, 1, 6)  # weather range 1-6
                ])
                n_trig, _, _ = execute_validation_rules(neighbor[0], neighbor[1], neighbor[2])
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = TARGET_PATHS[target_path_idx]
    samples = []
    attempts = 0

    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(1, 100),
            random.randint(1, 100),
            random.randint(1, 6)
        ])
        triggered, _, _ = execute_validation_rules(state[0], state[1], state[2])

        if not triggered:
            continue

        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)

        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }

        samples.append(sample_data)

    return samples


def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample[
                'robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)

        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]

    return selected_samples


def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = TARGET_PATHS[target_path_idx]
        samples = []
        attempts = 0

        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(1, 100),
                random.randint(1, 100),
                random.randint(1, 6)
            ])
            triggered, _, _ = execute_validation_rules(state[0], state[1], state[2])

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)

            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)

        random.shuffle(samples)
        return samples[:config.top_k_samples]

    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")

        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)


def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}

    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates

    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}

        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config,
                                                         shared_candidates[path_idx])

            strategy_results[path_idx] = samples

        results[strategy_name] = strategy_results

    return results


def analyze_fitness_values(results, config):
    analysis_results = {}

    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []

        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])

        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }

        all_scores = []

        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])

        scores_array = np.array(all_scores)

        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)

        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })

        analysis_results[strategy_name] = analysis

    return analysis_results


def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []

    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)

    df = pd.DataFrame(df_data)
    return df, analysis_results


def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []

    print(f"Starting {num_runs} experiments...")

    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")

        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)

        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)

        print(f"Experiment {run_idx + 1} completed.")

    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df


def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')

    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width

    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []

    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]

        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)

    stats_df = pd.DataFrame(stats_data)

    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Results saved to: {output_path}")


def main():
    results_df = run_multiple_experiments(num_runs=20)

    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")

    save_results_to_excel(results_df, output_path)

    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)

    return results_df, output_path


if __name__ == "__main__":
    results, output_file = main()