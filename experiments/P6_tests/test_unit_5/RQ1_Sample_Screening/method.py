import os
import random
from collections import deque
import numpy as np
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# --- 全局状态范围配置 ---
STATE_MIN_X, STATE_MAX_X = 1, 40
STATE_MIN_Y, STATE_MAX_Y = 1000, 10000
STATE_MIN_Z, STATE_MAX_Z = 800, 1500

def generate_input():
    return [
        random.randint(STATE_MIN_X, STATE_MAX_X),
        random.randint(STATE_MIN_Y, STATE_MAX_Y),
        random.randint(STATE_MIN_Z, STATE_MAX_Z)
    ]


def execute_Tr(light, co2, temp):
    """高CO2极端条件增强版分支测试函数"""
    actions = []
    triggered = set()

    # Fixed: removed extra space before if
    if (co2 > 1200) != (co2 > 400):
        triggered.add(1)
    if (co2 > 1200) != (co2 > 200):
        triggered.add(2)
    if (co2 > 1200) != (co2 > 100):
        triggered.add(3)
    if (co2 > 1200) != (co2 > 600):
        triggered.add(4)

    if (co2 > 1100 and light > 7000) != (co2 > 100 and light > 7000):
        triggered.add(5)
    if (co2 > 1100 and light > 7000) != (co2 > 500 and light > 7000):
        triggered.add(6)
    if (co2 > 1100 and light > 7000) != (co2 > 800 and light > 7000):
        triggered.add(7)
    if (co2 > 1100 and light > 7000) != (co2 > 1100 and light > 5000):
        triggered.add(8)
    if (co2 > 1100 and light > 7000) != (co2 > 1100 and light > 4000):
        triggered.add(9)
    if (co2 > 1100 and light > 7000) != (co2 > 1100 and light > 3000):
        triggered.add(10)
    if (co2 > 1100 and light > 7000) != (co2 > 1100 and light > 2000):
        triggered.add(11)

    if (co2 > 1250 and temp > 28) != (co2 > 950 and temp > 28):
        triggered.add(12)
    if (co2 > 1250 and temp > 28) != (co2 > 1250 and temp > 18):
        triggered.add(13)
    if (co2 > 1250 and temp > 28) != (co2 > 1000 and temp > 28):
        triggered.add(14)
    if (co2 > 1250 and temp > 28) != (co2 > 900 and temp > 28):
        triggered.add(15)
    if (co2 > 1250 and temp > 28) != (co2 > 1250 or temp > 28):
        triggered.add(16)
    if (co2 > 1250 and temp > 28) != (co2 > 1250 and temp <= 28):
        triggered.add(17)
    if (co2 > 1250 and temp > 28) != (co2 < 1250 and temp > 28):
        triggered.add(18)
    if (co2 > 1250 and temp > 28) != (light > 1250 and temp > 28):
        triggered.add(19)

    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1350 and light > 6000 and temp > 22):
        triggered.add(20)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 < 1150 and light > 6000 and temp > 22):
        triggered.add(21)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 7000 and temp > 22):
        triggered.add(22)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 8000 and temp > 22):
        triggered.add(23)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 4000 and temp > 22):
        triggered.add(24)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 6000 and temp > 32):
        triggered.add(25)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 6000 and temp > 12):
        triggered.add(26)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light < 6000 and temp > 22):
        triggered.add(27)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 6000 and temp < 22):
        triggered.add(28)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 or light > 6000 and temp > 22):
        triggered.add(29)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 6000 or temp > 22):
        triggered.add(30)

    if (co2 < 1050 and light > 6000) != (co2 < 1250 and light > 6000):
        triggered.add(31)
    if (co2 < 1050 and light > 6000) != (co2 < 1350 and light > 6000):
        triggered.add(32)
    if (co2 < 1050 and light > 6000) != (co2 < 1450 and light > 6000):
        triggered.add(33)
    if (co2 < 1050 and light > 6000) != (co2 > 1050 and light > 6000):
        triggered.add(34)
    if (co2 < 1050 and light > 6000) != (co2 < 1050 or light > 6000):
        triggered.add(35)
    if (co2 < 1050 and light > 6000) != (co2 < 1050 and light < 6000):
        triggered.add(36)
    if (co2 < 1050 and light > 6000) != (co2 < 1050 and light > 4000):
        triggered.add(37)
    if (co2 < 1050 and light > 6000) != (co2 < 1050 and light > 8000):
        triggered.add(38)

    if (temp > 28 and light > 6000) != (temp > 18 and light > 6000):
        triggered.add(39)
    if (temp > 28 and light > 6000) != (temp > 24 and light > 6000):
        triggered.add(40)
    if (temp > 28 and light > 6000) != (temp > 20 and light > 6000):
        triggered.add(41)
    if (temp > 28 and light > 6000) != (temp > 28 or light > 6000):
        triggered.add(42)
    if (temp > 28 and light > 6000) != (temp > 28 and light > 4000):
        triggered.add(43)
    if (temp > 28 and light > 6000) != (temp > 28 and light > 2000):
        triggered.add(44)
    if (temp > 28 and light > 6000) != (temp > 28 and light > 8000):
        triggered.add(45)

    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1300 and light > 6000 and temp > 22):
        triggered.add(46)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 600 and light > 6000 and temp > 22):
        triggered.add(47)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 300 and light > 6000 and temp > 22):
        triggered.add(48)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 < 1100 and light > 6000 and temp > 22):
        triggered.add(49)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1100 or light > 6000 and temp > 22):
        triggered.add(50)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1100 and light > 4000 and temp > 22):
        triggered.add(51)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1100 and light > 8000 and temp > 22):
        triggered.add(52)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1100 and light > 6000 or temp > 22):
        triggered.add(53)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1100 and light > 6000 and temp > 12):
        triggered.add(54)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1100 and light > 6000 and temp > 32):
        triggered.add(55)

    if (co2 < 1050 and light > 5000) != (co2 < 1250 and light > 5000):
        triggered.add(56)
    if (co2 < 1050 and light > 5000) != (co2 < 950 and light > 5000):
        triggered.add(57)
    if (co2 < 1050 and light > 5000) != (co2 < 1050 or light > 5000):
        triggered.add(58)
    if (co2 < 1050 and light > 5000) != (co2 < 1050 and light > 6000):
        triggered.add(59)
    if (co2 < 1050 and light > 5000) != (co2 < 1050 and light > 8000):
        triggered.add(60)
    if (co2 < 1050 and light > 5000) != (co2 < 1050 and light > 3000):
        triggered.add(61)

    if (3000 <= light <= 8000) != (2000 <= light <= 8000):
        triggered.add(62)
    if (3000 <= light <= 8000) != (1000 <= light <= 8000):
        triggered.add(63)
    if (3000 <= light <= 8000) != (4000 <= light <= 8000):
        triggered.add(64)
    if (3000 <= light <= 8000) != (6000 <= light <= 8000):
        triggered.add(65)

    if (co2 > 1180 and light > 5500) != (co2 > 1280 and light > 5500):
        triggered.add(66)
    if (co2 > 1180 and light > 5500) != (co2 > 1380 and light > 5500):
        triggered.add(67)
    if (co2 > 1180 and light > 5500) != (co2 < 1180 and light > 5500):
        triggered.add(68)
    if (co2 > 1180 and light > 5500) != (co2 > 1180 or light > 5500):
        triggered.add(69)
    if (co2 > 1180 and light > 5500) != (co2 > 1180 and light > 6600):
        triggered.add(70)
    if (co2 > 1180 and light > 5500) != (co2 > 1180 and light > 7700):
        triggered.add(71)
    if (co2 > 1180 and light > 5500) != (co2 > 1180 and light > 8800):
        triggered.add(72)

    if (light > 3500 and light < 8500) != (light > 5500 and light < 8500):
        triggered.add(73)
    if (light > 3500 and light < 8500) != (light > 6500 and light < 8500):
        triggered.add(74)
    if (light > 3500 and light < 8500) != (light > 7500 and light < 8500):
        triggered.add(75)
    if (light > 3500 and light < 8500) != (light < 3500 and light < 8500):
        triggered.add(76)
    if (light > 3500 and light < 8500) != (light > 3500 and light > 8500):
        triggered.add(77)
    if (light > 3500 and light < 8500) != (light > 3500 and light < 7500):
        triggered.add(78)
    if (light > 3500 and light < 8500) != (light > 3500 and light < 5500):
        triggered.add(79)
    if (light > 3500 and light < 8500) != (light > 3500 and light < 4500):
        triggered.add(80)

    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 900) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000):
        triggered.add(81)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1250) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000):
        triggered.add(82)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 3 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000):
        triggered.add(83)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 4500) ** 2 / 100 + (temp - 25) ** 2 < 10000):
        triggered.add(84)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 3500) ** 2 / 100 + (temp - 25) ** 2 < 10000):
        triggered.add(85)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 5500) ** 3 / 100 + (temp - 25) ** 2 < 10000):
        triggered.add(86)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 5500) ** 2 / 200 + (temp - 25) ** 2 < 10000):
        triggered.add(87)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 15) ** 2 < 10000):
        triggered.add(88)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 35) ** 2 < 10000):
        triggered.add(89)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 3 < 10000):
        triggered.add(90)

    if (light / temp > 50 and light / temp < 300) != (light / temp > 60 and light / temp < 300):
        triggered.add(91)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 70 and light / temp < 300):
        triggered.add(92)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 80 and light / temp < 300):
        triggered.add(93)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 90 and light / temp < 300):
        triggered.add(94)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 50 or light / temp < 300):
        triggered.add(95)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 50 and light / temp < 200):
        triggered.add(96)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 50 and light / temp < 100):
        triggered.add(97)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 50 and light / temp < 350):
        triggered.add(98)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 50 and light / temp < 400):
        triggered.add(99)

    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 1050) + abs(light - 5500) / 10 + abs(temp - 25) < 500):
        triggered.add(100)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 950) + abs(light - 5500) / 10 + abs(temp - 25) < 500):
        triggered.add(101)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 850) + abs(light - 5500) / 10 + abs(temp - 25) < 500):
        triggered.add(102)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 1150) + abs(light - 6500) / 10 + abs(temp - 25) < 500):
        triggered.add(103)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 1150) + abs(light - 7500) / 10 + abs(temp - 25) < 500):
        triggered.add(104)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 1150) + abs(light - 8500) / 20 + abs(temp - 25) < 500):
        triggered.add(105)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 15) < 500):
        triggered.add(106)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 400):
        triggered.add(107)

    if (light > 1200 and light < 9800) != (light > 2200 and light < 9800):
        triggered.add(108)
    if (light > 1200 and light < 9800) != (light > 4200 and light < 9800):
        triggered.add(109)
    if (light > 1200 and light < 9800) != (light > 6200 and light < 9800):
        triggered.add(110)
    if (light > 1200 and light < 9800) != (light > 7200 and light < 9800):
        triggered.add(111)
    if (light > 1200 and light < 9800) != (light > 1200 and light < 7800):
        triggered.add(112)
    if (light > 1200 and light < 9800) != (light > 1200 and light < 5800):
        triggered.add(113)
    if (light > 1200 and light < 9800) != (light > 1200 and light < 2800):
        triggered.add(114)
    if (light > 1200 and light < 9800) != (light > 1200 and light < 3800):
        triggered.add(115)
    if (light > 1200 and light < 9800) != (light > 1200 and light < 4800):
        triggered.add(116)

    if (temp > 3 and temp < 38) != (temp > 13 and temp < 38):
        triggered.add(117)
    if (temp > 3 and temp < 38) != (temp > 23 and temp < 38):
        triggered.add(118)
    if (temp > 3 and temp < 38) != (temp > 33 and temp < 38):
        triggered.add(119)
    if (temp > 3 and temp < 38) != (temp > 3 and temp < 8):
        triggered.add(120)
    if (temp > 3 and temp < 38) != (temp > 3 and temp < 18):
        triggered.add(121)
    if (temp > 3 and temp < 38) != (temp > 3 and temp < 28):
        triggered.add(122)
    if (temp > 3 and temp < 38) != (temp > 3 and temp < 30):
        triggered.add(123)

    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 1000 and light > 1500 and temp > 5):
        triggered.add(124)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 950 and light > 1500 and temp > 5):
        triggered.add(125)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 1250 and light > 1500 and temp > 5):
        triggered.add(126)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 or light > 1500 and temp > 5):
        triggered.add(127)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 3500 and temp > 5):
        triggered.add(128)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 5500 and temp > 5):
        triggered.add(129)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 6500 and temp > 5):
        triggered.add(130)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 1500 and temp > 15):
        triggered.add(131)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 1500 and temp > 25):
        triggered.add(132)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 1500 and temp > 35):
        triggered.add(133)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 1500 or temp > 5):
        triggered.add(134)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 2500 and temp > 5):
        triggered.add(135)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 4500 and temp > 5):
        triggered.add(136)

    return triggered


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0


# Fixed: Proper Python list syntax (removed A1 =, A2 =, etc.)
TARGET_PATHS = [
    [1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20, 21, 22, 23, 25, 27, 28, 31, 32, 33, 34, 35, 45, 46, 49,
     52, 55, 56, 58, 66, 67, 68, 70, 71, 72, 74, 75, 76, 77, 79, 80, 81, 82, 83, 84, 85, 86, 96, 97, 111, 113, 114,
     115, 116, 119, 120, 121, 122, 123, 126, 130, 133],
    [1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20, 21, 22, 23, 25, 27, 28, 31, 32, 33, 34, 35, 45, 46, 49, 52,
     55, 56, 58, 66, 67, 68, 70, 71, 72, 74, 75, 76, 77, 79, 80, 81, 83, 84, 85, 86, 97, 110, 111, 113, 114, 115, 116,
     119, 120, 121, 122, 123, 126, 130, 133],
    [1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20, 21, 22, 23, 25, 27, 28, 31, 32, 33, 34, 35, 45, 46, 49, 52,
     55, 56, 58, 68, 69, 74, 75, 76, 77, 79, 80, 81, 82, 84, 85, 86, 96, 97, 110, 111, 113, 114, 115, 116, 119, 120,
     121, 122, 123, 126, 130, 133],
    [8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20, 21, 22, 23, 27, 28, 31, 32, 33, 34, 35, 45, 46, 49, 52, 56, 58, 66, 67,
     68, 70, 71, 72, 74, 75, 76, 77, 79, 80, 81, 83, 84, 85, 86, 90, 97, 110, 111, 113, 114, 115, 116, 120, 121, 122,
     123, 126, 130],
    [1, 2, 3, 4, 8, 9, 10, 11, 26, 28, 29, 30, 31, 32, 33, 34, 35, 39, 41, 42, 50, 53, 54, 56, 58, 66, 67, 68, 70, 71,
     72, 74, 75, 76, 77, 79, 80, 81, 83, 84, 85, 86, 96, 97, 110, 111, 113, 114, 115, 116, 118, 119, 120, 121, 126,
     130, 132, 133],
    [1, 2, 3, 4, 8, 9, 10, 11, 20, 21, 22, 23, 25, 27, 28, 31, 32, 33, 34, 35, 39, 40, 41, 42, 46, 49, 52, 55, 56, 58,
     68, 69, 74, 75, 76, 77, 79, 80, 81, 82, 84, 85, 86, 96, 97, 110, 111, 113, 114, 115, 116, 119, 120, 121, 126, 130,
     133],
    [1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 24, 27, 29, 30, 42, 43, 44, 50, 51, 53, 56, 58, 65, 66, 67, 68,
     70, 71, 72, 74, 75, 76, 77, 79, 80, 81, 83, 84, 85, 86, 97, 110, 111, 113, 114, 115, 116, 120, 121, 122, 123, 126,
     130],
    [1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20, 21, 22, 23, 25, 27, 28, 31, 32, 33, 34, 35, 45, 46, 49, 52,
     55, 56, 58, 66, 67, 68, 71, 72, 75, 76, 77, 79, 80, 87, 96, 97, 111, 113, 114, 115, 116, 119, 120, 121, 122, 126,
     133],
    [1, 2, 3, 4, 8, 9, 10, 11, 26, 28, 29, 30, 31, 32, 33, 34, 35, 39, 42, 50, 53, 54, 56, 58, 68, 69, 74, 75, 76, 77,
     79, 80, 81, 82, 83, 84, 85, 86, 95, 98, 99, 110, 111, 113, 114, 115, 116, 118, 119, 120, 121, 126, 130, 132, 133],
    [1, 2, 3, 4, 8, 9, 10, 11, 26, 28, 29, 30, 31, 32, 33, 34, 35, 42, 50, 53, 54, 56, 58, 66, 67, 68, 70, 71, 72, 74,
     75, 76, 77, 79, 80, 81, 82, 83, 84, 85, 86, 95, 111, 113, 114, 115, 116, 118, 119, 120, 126, 130, 131, 132, 133],
    [1, 2, 3, 4, 12, 14, 15, 16, 18, 19, 21, 29, 30, 31, 32, 33, 34, 35, 45, 47, 48, 49, 50, 53, 56, 58, 68, 69, 74,
     75, 76, 77, 79, 80, 81, 82, 84, 85, 86, 97, 110, 111, 113, 114, 115, 116, 119, 120, 121, 122, 123, 126, 130, 133],
    [8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20, 21, 22, 23, 25, 27, 28, 31, 32, 33, 34, 35, 45, 46, 49, 52, 55, 56, 58,
     66, 67, 68, 70, 71, 72, 75, 76, 77, 79, 80, 96, 97, 102, 111, 113, 114, 115, 116, 119, 120, 121, 122, 126, 133],
    [1, 2, 3, 4, 8, 9, 10, 11, 28, 29, 30, 31, 32, 33, 34, 35, 42, 50, 53, 56, 58, 66, 67, 68, 70, 71, 72, 74, 75, 76,
     77, 79, 80, 81, 83, 84, 85, 86, 95, 110, 111, 113, 114, 115, 116, 117, 118, 119, 120, 126, 130, 131, 132, 133],
    [1, 2, 3, 4, 5, 6, 7, 16, 18, 19, 21, 29, 30, 34, 36, 38, 45, 47, 48, 49, 50, 53, 60, 68, 69, 76, 77, 78, 79, 80,
     83, 96, 97, 100, 101, 102, 103, 104, 105, 112, 113, 114, 115, 116, 119, 120, 121, 122, 123, 124, 125, 126, 133],
    [1, 2, 3, 4, 12, 14, 15, 16, 18, 19, 21, 29, 30, 31, 32, 33, 34, 35, 45, 47, 48, 49, 50, 53, 56, 58, 68, 69, 74,
     75, 76, 77, 79, 80, 81, 82, 84, 85, 86, 88, 90, 97, 111, 113, 114, 115, 116, 120, 121, 122, 123, 126, 130, 133],
    [8, 9, 10, 11, 13, 16, 17, 20, 21, 22, 23, 25, 27, 28, 32, 33, 34, 35, 39, 41, 42, 46, 49, 52, 55, 58, 67, 68, 70,
     71, 72, 74, 75, 76, 77, 79, 80, 82, 96, 97, 102, 110, 111, 113, 114, 115, 116, 119, 120, 121, 130, 132, 133],
    [8, 9, 10, 11, 26, 28, 29, 30, 31, 32, 33, 34, 35, 42, 50, 53, 54, 56, 58, 66, 67, 68, 70, 71, 72, 74, 75, 76, 77,
     79, 80, 81, 83, 84, 85, 86, 89, 95, 99, 111, 113, 114, 115, 116, 118, 119, 120, 121, 126, 130, 132, 133],
    [8, 9, 10, 11, 13, 16, 17, 21, 22, 23, 25, 27, 28, 34, 35, 39, 41, 42, 49, 52, 55, 58, 68, 70, 71, 72, 74, 75, 76,
     77, 79, 80, 96, 97, 100, 101, 102, 107, 111, 113, 114, 115, 116, 118, 119, 120, 121, 130, 132, 133],
    [1, 2, 3, 4, 5, 6, 7, 16, 18, 19, 21, 29, 30, 34, 36, 38, 45, 47, 48, 49, 50, 53, 60, 68, 69, 75, 76, 77, 79, 80,
     83, 96, 97, 100, 101, 102, 103, 104, 105, 113, 114, 115, 116, 119, 120, 121, 122, 123, 127, 134],
    [1, 2, 3, 4, 5, 6, 7, 15, 16, 18, 19, 21, 29, 30, 34, 36, 38, 45, 47, 48, 49, 50, 53, 60, 68, 69, 76, 77, 78, 79,
     80, 83, 96, 97, 106, 107, 112, 113, 114, 115, 116, 119, 120, 121, 122, 123, 124, 125, 126, 133],
    [1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 24, 27, 29, 30, 42, 43, 44, 50, 51, 53, 56, 58, 65, 73, 74, 75,
     76, 77, 80, 81, 85, 97, 110, 111, 114, 115, 116, 119, 120, 121, 122, 123, 126, 129, 130, 133],
    [1, 2, 3, 4, 12, 15, 16, 18, 19, 21, 29, 30, 34, 36, 38, 45, 47, 48, 49, 50, 53, 57, 60, 68, 69, 74, 75, 76, 77,
     79, 80, 81, 83, 96, 97, 110, 111, 113, 114, 115, 116, 119, 120, 121, 122, 124, 126, 130, 133],
    [1, 2, 3, 4, 9, 10, 11, 12, 14, 15, 16, 18, 19, 24, 27, 29, 30, 42, 43, 44, 50, 51, 53, 65, 69, 73, 74, 75, 76,
     77, 84, 85, 86, 87, 97, 110, 111, 114, 115, 119, 120, 121, 122, 123, 126, 129, 130, 133, 136],
    [1, 2, 3, 4, 9, 10, 11, 12, 14, 15, 16, 18, 19, 24, 27, 29, 30, 42, 43, 44, 50, 51, 53, 65, 73, 74, 75, 76, 77,
     84, 85, 86, 97, 109, 110, 111, 114, 115, 119, 120, 121, 122, 123, 126, 129, 130, 133, 136],
    [1, 2, 3, 4, 12, 14, 15, 16, 18, 19, 27, 29, 30, 42, 50, 53, 63, 69, 76, 86, 91, 92, 93, 94, 100, 101, 102, 103,
     104, 107, 108, 109, 110, 111, 119, 120, 121, 122, 123, 126, 128, 129, 130, 133, 135, 136],
    [1, 2, 3, 4, 16, 18, 19, 30, 35, 36, 37, 42, 43, 44, 53, 58, 61, 65, 73, 74, 75, 76, 77, 83, 86, 97, 103, 104,
     105, 107, 110, 111, 114, 115, 119, 120, 121, 122, 123, 124, 125, 126, 129, 130, 133, 136],
    [1, 2, 3, 4, 12, 15, 16, 18, 19, 30, 35, 36, 42, 44, 53, 58, 62, 63, 76, 83, 86, 92, 93, 94, 100, 101, 102, 105,
     108, 109, 110, 111, 119, 120, 121, 122, 123, 124, 126, 128, 129, 130, 133, 135, 136],
    [1, 2, 3, 4, 16, 18, 19, 30, 35, 36, 42, 44, 53, 58, 61, 64, 65, 73, 74, 75, 76, 77, 83, 86, 97, 103, 104, 105,
     107, 109, 110, 111, 114, 115, 120, 121, 122, 123, 124, 125, 126, 129, 130, 133, 136],
    [1, 2, 3, 4, 15, 16, 18, 19, 30, 35, 36, 37, 42, 43, 44, 53, 59, 60, 65, 68, 69, 74, 75, 76, 77, 79, 80, 81, 83,
     96, 97, 110, 111, 113, 114, 115, 116, 119, 120, 121, 122, 124, 125, 126, 130, 133]
]


class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 2000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(TARGET_PATHS)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }


def compute_robustness(state, path):
    base = execute_Tr(state[0], state[1], state[2])
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                    np.clip(state[0] + dx, STATE_MIN_X, STATE_MAX_X),
                    np.clip(state[1] + dy, STATE_MIN_Y, STATE_MAX_Y),
                    np.clip(state[2] + dz, STATE_MIN_Z, STATE_MAX_Z)
                ])
                n_trig = execute_Tr(neighbor[0], neighbor[1], neighbor[2])
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = TARGET_PATHS[target_path_idx]
    samples = []
    attempts = 0

    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(STATE_MIN_X, STATE_MAX_X),
            random.randint(STATE_MIN_Y, STATE_MAX_Y),
            random.randint(STATE_MIN_Z, STATE_MAX_Z)
        ])
        triggered = execute_Tr(state[0], state[1], state[2])

        if not triggered:
            continue

        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)

        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }

        samples.append(sample_data)

    return samples


def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample[
                'robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)

        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]

    return selected_samples


def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = TARGET_PATHS[target_path_idx]
        samples = []
        attempts = 0

        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(STATE_MIN_X, STATE_MAX_X),
                random.randint(STATE_MIN_Y, STATE_MAX_Y),
                random.randint(STATE_MIN_Z, STATE_MAX_Z)
            ])
            triggered = execute_Tr(state[0], state[1], state[2])

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)

            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)

        random.shuffle(samples)
        return samples[:config.top_k_samples]

    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")

        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)


def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}

    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates

    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}

        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config,
                                                         shared_candidates[path_idx])

            strategy_results[path_idx] = samples

        results[strategy_name] = strategy_results

    return results


def analyze_fitness_values(results, config):
    analysis_results = {}

    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []

        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])

        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }

        all_scores = []

        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])

        scores_array = np.array(all_scores)

        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)

        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })

        analysis_results[strategy_name] = analysis

    return analysis_results


def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []

    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)

    df = pd.DataFrame(df_data)
    return df, analysis_results


def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []

    print(f"Starting {num_runs} experiments...")

    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")

        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)

        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)

        print(f"Experiment {run_idx + 1} completed.")

    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df


def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')

    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width

    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []

    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]

        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)

    stats_df = pd.DataFrame(stats_data)

    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Results saved to: {output_path}")


def main():
    results_df = run_multiple_experiments(num_runs=20)

    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")

    save_results_to_excel(results_df, output_path)

    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)

    return results_df, output_path


if __name__ == "__main__":
    results, output_file = main()