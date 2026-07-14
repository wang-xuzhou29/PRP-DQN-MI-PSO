
import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
from torch.distributions import Normal
import numpy as np
import random
import time
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# ===  ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'ACTION_DIM': 3,
    # : [co2, moisture, temp]
    'MIN_VALUES': np.array([800, 10, 1], dtype=np.float32),
    'MAX_VALUES': np.array([1500, 80, 40], dtype=np.float32),
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 3,
    'REPLAY_BATCH_SIZE': 64,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'ACTOR_LR': 3e-4,
    'CRITIC_LR': 3e-4,
    'ALPHA_LR': 3e-4,
    'GAMMA': 0.99,
    'TAU': 0.005,
    'NUM_RUNS': 20,
    'TOP_K_SAMPLES': 20,
    'TARGET_PATHS': [
        [1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20, 21, 22, 23, 25, 27, 28, 31, 32, 33, 34, 35, 45, 46, 49,
         52, 55, 56, 58, 66, 67, 68, 70, 71, 72, 74, 75, 76, 77, 79, 80, 81, 82, 83, 84, 85, 86, 96, 97, 111, 113, 114,
         115, 116, 119, 120, 121, 122, 123, 126, 130, 133],
        [1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20, 21, 22, 23, 25, 27, 28, 31, 32, 33, 34, 35, 45, 46, 49,
         52,
         55, 56, 58, 66, 67, 68, 70, 71, 72, 74, 75, 76, 77, 79, 80, 81, 83, 84, 85, 86, 97, 110, 111, 113, 114, 115,
         116,
         119, 120, 121, 122, 123, 126, 130, 133],
        [1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20, 21, 22, 23, 25, 27, 28, 31, 32, 33, 34, 35, 45, 46, 49,
         52,
         55, 56, 58, 68, 69, 74, 75, 76, 77, 79, 80, 81, 82, 84, 85, 86, 96, 97, 110, 111, 113, 114, 115, 116, 119, 120,
         121, 122, 123, 126, 130, 133],
        [8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20, 21, 22, 23, 27, 28, 31, 32, 33, 34, 35, 45, 46, 49, 52, 56, 58, 66,
         67,
         68, 70, 71, 72, 74, 75, 76, 77, 79, 80, 81, 83, 84, 85, 86, 90, 97, 110, 111, 113, 114, 115, 116, 120, 121,
         122,
         123, 126, 130],
        [1, 2, 3, 4, 8, 9, 10, 11, 26, 28, 29, 30, 31, 32, 33, 34, 35, 39, 41, 42, 50, 53, 54, 56, 58, 66, 67, 68, 70,
         71,
         72, 74, 75, 76, 77, 79, 80, 81, 83, 84, 85, 86, 96, 97, 110, 111, 113, 114, 115, 116, 118, 119, 120, 121, 126,
         130, 132, 133],
        [1, 2, 3, 4, 8, 9, 10, 11, 20, 21, 22, 23, 25, 27, 28, 31, 32, 33, 34, 35, 39, 40, 41, 42, 46, 49, 52, 55, 56,
         58,
         68, 69, 74, 75, 76, 77, 79, 80, 81, 82, 84, 85, 86, 96, 97, 110, 111, 113, 114, 115, 116, 119, 120, 121, 126,
         130,
         133],
        [1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 24, 27, 29, 30, 42, 43, 44, 50, 51, 53, 56, 58, 65, 66, 67,
         68,
         70, 71, 72, 74, 75, 76, 77, 79, 80, 81, 83, 84, 85, 86, 97, 110, 111, 113, 114, 115, 116, 120, 121, 122, 123,
         126,
         130],
        [1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20, 21, 22, 23, 25, 27, 28, 31, 32, 33, 34, 35, 45, 46, 49,
         52,
         55, 56, 58, 66, 67, 68, 71, 72, 75, 76, 77, 79, 80, 87, 96, 97, 111, 113, 114, 115, 116, 119, 120, 121, 122,
         126,
         133],
        [1, 2, 3, 4, 8, 9, 10, 11, 26, 28, 29, 30, 31, 32, 33, 34, 35, 39, 42, 50, 53, 54, 56, 58, 68, 69, 74, 75, 76,
         77,
         79, 80, 81, 82, 83, 84, 85, 86, 95, 98, 99, 110, 111, 113, 114, 115, 116, 118, 119, 120, 121, 126, 130, 132,
         133],
        [1, 2, 3, 4, 8, 9, 10, 11, 26, 28, 29, 30, 31, 32, 33, 34, 35, 42, 50, 53, 54, 56, 58, 66, 67, 68, 70, 71, 72,
         74,
         75, 76, 77, 79, 80, 81, 82, 83, 84, 85, 86, 95, 111, 113, 114, 115, 116, 118, 119, 120, 126, 130, 131, 132,
         133],
        [1, 2, 3, 4, 12, 14, 15, 16, 18, 19, 21, 29, 30, 31, 32, 33, 34, 35, 45, 47, 48, 49, 50, 53, 56, 58, 68, 69, 74,
         75, 76, 77, 79, 80, 81, 82, 84, 85, 86, 97, 110, 111, 113, 114, 115, 116, 119, 120, 121, 122, 123, 126, 130,
         133],
        [8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20, 21, 22, 23, 25, 27, 28, 31, 32, 33, 34, 35, 45, 46, 49, 52, 55, 56,
         58,
         66, 67, 68, 70, 71, 72, 75, 76, 77, 79, 80, 96, 97, 102, 111, 113, 114, 115, 116, 119, 120, 121, 122, 126,
         133],
        [1, 2, 3, 4, 8, 9, 10, 11, 28, 29, 30, 31, 32, 33, 34, 35, 42, 50, 53, 56, 58, 66, 67, 68, 70, 71, 72, 74, 75,
         76,
         77, 79, 80, 81, 83, 84, 85, 86, 95, 110, 111, 113, 114, 115, 116, 117, 118, 119, 120, 126, 130, 131, 132, 133],
        [1, 2, 3, 4, 5, 6, 7, 16, 18, 19, 21, 29, 30, 34, 36, 38, 45, 47, 48, 49, 50, 53, 60, 68, 69, 76, 77, 78, 79,
         80,
         83, 96, 97, 100, 101, 102, 103, 104, 105, 112, 113, 114, 115, 116, 119, 120, 121, 122, 123, 124, 125, 126,
         133],
        [1, 2, 3, 4, 12, 14, 15, 16, 18, 19, 21, 29, 30, 31, 32, 33, 34, 35, 45, 47, 48, 49, 50, 53, 56, 58, 68, 69, 74,
         75, 76, 77, 79, 80, 81, 82, 84, 85, 86, 88, 90, 97, 111, 113, 114, 115, 116, 120, 121, 122, 123, 126, 130,
         133],
        [8, 9, 10, 11, 13, 16, 17, 20, 21, 22, 23, 25, 27, 28, 32, 33, 34, 35, 39, 41, 42, 46, 49, 52, 55, 58, 67, 68,
         70,
         71, 72, 74, 75, 76, 77, 79, 80, 82, 96, 97, 102, 110, 111, 113, 114, 115, 116, 119, 120, 121, 130, 132, 133],
        [8, 9, 10, 11, 26, 28, 29, 30, 31, 32, 33, 34, 35, 42, 50, 53, 54, 56, 58, 66, 67, 68, 70, 71, 72, 74, 75, 76,
         77,
         79, 80, 81, 83, 84, 85, 86, 89, 95, 99, 111, 113, 114, 115, 116, 118, 119, 120, 121, 126, 130, 132, 133],
        [8, 9, 10, 11, 13, 16, 17, 21, 22, 23, 25, 27, 28, 34, 35, 39, 41, 42, 49, 52, 55, 58, 68, 70, 71, 72, 74, 75,
         76,
         77, 79, 80, 96, 97, 100, 101, 102, 107, 111, 113, 114, 115, 116, 118, 119, 120, 121, 130, 132, 133],
        [1, 2, 3, 4, 5, 6, 7, 16, 18, 19, 21, 29, 30, 34, 36, 38, 45, 47, 48, 49, 50, 53, 60, 68, 69, 75, 76, 77, 79,
         80,
         83, 96, 97, 100, 101, 102, 103, 104, 105, 113, 114, 115, 116, 119, 120, 121, 122, 123, 127, 134],
        [1, 2, 3, 4, 5, 6, 7, 15, 16, 18, 19, 21, 29, 30, 34, 36, 38, 45, 47, 48, 49, 50, 53, 60, 68, 69, 76, 77, 78,
         79,
         80, 83, 96, 97, 106, 107, 112, 113, 114, 115, 116, 119, 120, 121, 122, 123, 124, 125, 126, 133],
        [1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 24, 27, 29, 30, 42, 43, 44, 50, 51, 53, 56, 58, 65, 73, 74,
         75,
         76, 77, 80, 81, 85, 97, 110, 111, 114, 115, 116, 119, 120, 121, 122, 123, 126, 129, 130, 133],
        [1, 2, 3, 4, 12, 15, 16, 18, 19, 21, 29, 30, 34, 36, 38, 45, 47, 48, 49, 50, 53, 57, 60, 68, 69, 74, 75, 76, 77,
         79, 80, 81, 83, 96, 97, 110, 111, 113, 114, 115, 116, 119, 120, 121, 122, 124, 126, 130, 133],
        [1, 2, 3, 4, 9, 10, 11, 12, 14, 15, 16, 18, 19, 24, 27, 29, 30, 42, 43, 44, 50, 51, 53, 65, 69, 73, 74, 75, 76,
         77, 84, 85, 86, 87, 97, 110, 111, 114, 115, 119, 120, 121, 122, 123, 126, 129, 130, 133, 136],
        [1, 2, 3, 4, 9, 10, 11, 12, 14, 15, 16, 18, 19, 24, 27, 29, 30, 42, 43, 44, 50, 51, 53, 65, 73, 74, 75, 76, 77,
         84, 85, 86, 97, 109, 110, 111, 114, 115, 119, 120, 121, 122, 123, 126, 129, 130, 133, 136],
        [1, 2, 3, 4, 12, 14, 15, 16, 18, 19, 27, 29, 30, 42, 50, 53, 63, 69, 76, 86, 91, 92, 93, 94, 100, 101, 102, 103,
         104, 107, 108, 109, 110, 111, 119, 120, 121, 122, 123, 126, 128, 129, 130, 133, 135, 136],
        [1, 2, 3, 4, 16, 18, 19, 30, 35, 36, 37, 42, 43, 44, 53, 58, 61, 65, 73, 74, 75, 76, 77, 83, 86, 97, 103, 104,
         105, 107, 110, 111, 114, 115, 119, 120, 121, 122, 123, 124, 125, 126, 129, 130, 133, 136],
        [1, 2, 3, 4, 12, 15, 16, 18, 19, 30, 35, 36, 42, 44, 53, 58, 62, 63, 76, 83, 86, 92, 93, 94, 100, 101, 102, 105,
         108, 109, 110, 111, 119, 120, 121, 122, 123, 124, 126, 128, 129, 130, 133, 135, 136],
        [1, 2, 3, 4, 16, 18, 19, 30, 35, 36, 42, 44, 53, 58, 61, 64, 65, 73, 74, 75, 76, 77, 83, 86, 97, 103, 104, 105,
         107, 109, 110, 111, 114, 115, 120, 121, 122, 123, 124, 125, 126, 129, 130, 133, 136],
        [1, 2, 3, 4, 15, 16, 18, 19, 30, 35, 36, 37, 42, 43, 44, 53, 59, 60, 65, 68, 69, 74, 75, 76, 77, 79, 80, 81, 83,
         96, 97, 110, 111, 113, 114, 115, 116, 119, 120, 121, 122, 124, 125, 126, 130, 133]
    ],
}


# ===  ===
def clip_state(state):
    """"""
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return np.clip(state, min_vals, max_vals)


def normalize_state(state):
    """[-1, 1]"""
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return 2 * (state - min_vals) / (max_vals - min_vals) - 1


def denormalize_state(normalized_state):
    """"""
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return (normalized_state + 1) * (max_vals - min_vals) / 2 + min_vals


def coverage_similarity(triggered, target_path):
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0

    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)


def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']

    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']

    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']

    return reward


def safe_divide(a, b):
    """, """
    return a / b if b != 0 else 0


def execute_Tr(light, co2, temp):
    """高CO2极端条件增强版分支测试函数"""
    actions = []
    triggered = set()

    # Fixed: removed extra space before if
    if (co2 > 1200) != (co2 > 400):
        triggered.add(1)
    if (co2 > 1200) != (co2 > 200):
        triggered.add(2)
    if (co2 > 1200) != (co2 > 100):
        triggered.add(3)
    if (co2 > 1200) != (co2 > 600):
        triggered.add(4)

    if (co2 > 1100 and light > 7000) != (co2 > 100 and light > 7000):
        triggered.add(5)
    if (co2 > 1100 and light > 7000) != (co2 > 500 and light > 7000):
        triggered.add(6)
    if (co2 > 1100 and light > 7000) != (co2 > 800 and light > 7000):
        triggered.add(7)
    if (co2 > 1100 and light > 7000) != (co2 > 1100 and light > 5000):
        triggered.add(8)
    if (co2 > 1100 and light > 7000) != (co2 > 1100 and light > 4000):
        triggered.add(9)
    if (co2 > 1100 and light > 7000) != (co2 > 1100 and light > 3000):
        triggered.add(10)
    if (co2 > 1100 and light > 7000) != (co2 > 1100 and light > 2000):
        triggered.add(11)

    if (co2 > 1250 and temp > 28) != (co2 > 950 and temp > 28):
        triggered.add(12)
    if (co2 > 1250 and temp > 28) != (co2 > 1250 and temp > 18):
        triggered.add(13)
    if (co2 > 1250 and temp > 28) != (co2 > 1000 and temp > 28):
        triggered.add(14)
    if (co2 > 1250 and temp > 28) != (co2 > 900 and temp > 28):
        triggered.add(15)
    if (co2 > 1250 and temp > 28) != (co2 > 1250 or temp > 28):
        triggered.add(16)
    if (co2 > 1250 and temp > 28) != (co2 > 1250 and temp <= 28):
        triggered.add(17)
    if (co2 > 1250 and temp > 28) != (co2 < 1250 and temp > 28):
        triggered.add(18)
    if (co2 > 1250 and temp > 28) != (light > 1250 and temp > 28):
        triggered.add(19)

    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1350 and light > 6000 and temp > 22):
        triggered.add(20)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 < 1150 and light > 6000 and temp > 22):
        triggered.add(21)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 7000 and temp > 22):
        triggered.add(22)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 8000 and temp > 22):
        triggered.add(23)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 4000 and temp > 22):
        triggered.add(24)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 6000 and temp > 32):
        triggered.add(25)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 6000 and temp > 12):
        triggered.add(26)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light < 6000 and temp > 22):
        triggered.add(27)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 6000 and temp < 22):
        triggered.add(28)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 or light > 6000 and temp > 22):
        triggered.add(29)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 6000 or temp > 22):
        triggered.add(30)

    if (co2 < 1050 and light > 6000) != (co2 < 1250 and light > 6000):
        triggered.add(31)
    if (co2 < 1050 and light > 6000) != (co2 < 1350 and light > 6000):
        triggered.add(32)
    if (co2 < 1050 and light > 6000) != (co2 < 1450 and light > 6000):
        triggered.add(33)
    if (co2 < 1050 and light > 6000) != (co2 > 1050 and light > 6000):
        triggered.add(34)
    if (co2 < 1050 and light > 6000) != (co2 < 1050 or light > 6000):
        triggered.add(35)
    if (co2 < 1050 and light > 6000) != (co2 < 1050 and light < 6000):
        triggered.add(36)
    if (co2 < 1050 and light > 6000) != (co2 < 1050 and light > 4000):
        triggered.add(37)
    if (co2 < 1050 and light > 6000) != (co2 < 1050 and light > 8000):
        triggered.add(38)

    if (temp > 28 and light > 6000) != (temp > 18 and light > 6000):
        triggered.add(39)
    if (temp > 28 and light > 6000) != (temp > 24 and light > 6000):
        triggered.add(40)
    if (temp > 28 and light > 6000) != (temp > 20 and light > 6000):
        triggered.add(41)
    if (temp > 28 and light > 6000) != (temp > 28 or light > 6000):
        triggered.add(42)
    if (temp > 28 and light > 6000) != (temp > 28 and light > 4000):
        triggered.add(43)
    if (temp > 28 and light > 6000) != (temp > 28 and light > 2000):
        triggered.add(44)
    if (temp > 28 and light > 6000) != (temp > 28 and light > 8000):
        triggered.add(45)

    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1300 and light > 6000 and temp > 22):
        triggered.add(46)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 600 and light > 6000 and temp > 22):
        triggered.add(47)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 300 and light > 6000 and temp > 22):
        triggered.add(48)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 < 1100 and light > 6000 and temp > 22):
        triggered.add(49)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1100 or light > 6000 and temp > 22):
        triggered.add(50)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1100 and light > 4000 and temp > 22):
        triggered.add(51)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1100 and light > 8000 and temp > 22):
        triggered.add(52)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1100 and light > 6000 or temp > 22):
        triggered.add(53)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1100 and light > 6000 and temp > 12):
        triggered.add(54)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1100 and light > 6000 and temp > 32):
        triggered.add(55)

    if (co2 < 1050 and light > 5000) != (co2 < 1250 and light > 5000):
        triggered.add(56)
    if (co2 < 1050 and light > 5000) != (co2 < 950 and light > 5000):
        triggered.add(57)
    if (co2 < 1050 and light > 5000) != (co2 < 1050 or light > 5000):
        triggered.add(58)
    if (co2 < 1050 and light > 5000) != (co2 < 1050 and light > 6000):
        triggered.add(59)
    if (co2 < 1050 and light > 5000) != (co2 < 1050 and light > 8000):
        triggered.add(60)
    if (co2 < 1050 and light > 5000) != (co2 < 1050 and light > 3000):
        triggered.add(61)

    if (3000 <= light <= 8000) != (2000 <= light <= 8000):
        triggered.add(62)
    if (3000 <= light <= 8000) != (1000 <= light <= 8000):
        triggered.add(63)
    if (3000 <= light <= 8000) != (4000 <= light <= 8000):
        triggered.add(64)
    if (3000 <= light <= 8000) != (6000 <= light <= 8000):
        triggered.add(65)

    if (co2 > 1180 and light > 5500) != (co2 > 1280 and light > 5500):
        triggered.add(66)
    if (co2 > 1180 and light > 5500) != (co2 > 1380 and light > 5500):
        triggered.add(67)
    if (co2 > 1180 and light > 5500) != (co2 < 1180 and light > 5500):
        triggered.add(68)
    if (co2 > 1180 and light > 5500) != (co2 > 1180 or light > 5500):
        triggered.add(69)
    if (co2 > 1180 and light > 5500) != (co2 > 1180 and light > 6600):
        triggered.add(70)
    if (co2 > 1180 and light > 5500) != (co2 > 1180 and light > 7700):
        triggered.add(71)
    if (co2 > 1180 and light > 5500) != (co2 > 1180 and light > 8800):
        triggered.add(72)

    if (light > 3500 and light < 8500) != (light > 5500 and light < 8500):
        triggered.add(73)
    if (light > 3500 and light < 8500) != (light > 6500 and light < 8500):
        triggered.add(74)
    if (light > 3500 and light < 8500) != (light > 7500 and light < 8500):
        triggered.add(75)
    if (light > 3500 and light < 8500) != (light < 3500 and light < 8500):
        triggered.add(76)
    if (light > 3500 and light < 8500) != (light > 3500 and light > 8500):
        triggered.add(77)
    if (light > 3500 and light < 8500) != (light > 3500 and light < 7500):
        triggered.add(78)
    if (light > 3500 and light < 8500) != (light > 3500 and light < 5500):
        triggered.add(79)
    if (light > 3500 and light < 8500) != (light > 3500 and light < 4500):
        triggered.add(80)

    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 900) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000):
        triggered.add(81)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1250) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000):
        triggered.add(82)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 3 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000):
        triggered.add(83)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 4500) ** 2 / 100 + (temp - 25) ** 2 < 10000):
        triggered.add(84)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 3500) ** 2 / 100 + (temp - 25) ** 2 < 10000):
        triggered.add(85)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 5500) ** 3 / 100 + (temp - 25) ** 2 < 10000):
        triggered.add(86)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 5500) ** 2 / 200 + (temp - 25) ** 2 < 10000):
        triggered.add(87)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 15) ** 2 < 10000):
        triggered.add(88)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 35) ** 2 < 10000):
        triggered.add(89)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 3 < 10000):
        triggered.add(90)

    if (light / temp > 50 and light / temp < 300) != (light / temp > 60 and light / temp < 300):
        triggered.add(91)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 70 and light / temp < 300):
        triggered.add(92)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 80 and light / temp < 300):
        triggered.add(93)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 90 and light / temp < 300):
        triggered.add(94)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 50 or light / temp < 300):
        triggered.add(95)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 50 and light / temp < 200):
        triggered.add(96)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 50 and light / temp < 100):
        triggered.add(97)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 50 and light / temp < 350):
        triggered.add(98)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 50 and light / temp < 400):
        triggered.add(99)

    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 1050) + abs(light - 5500) / 10 + abs(temp - 25) < 500):
        triggered.add(100)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 950) + abs(light - 5500) / 10 + abs(temp - 25) < 500):
        triggered.add(101)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 850) + abs(light - 5500) / 10 + abs(temp - 25) < 500):
        triggered.add(102)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 1150) + abs(light - 6500) / 10 + abs(temp - 25) < 500):
        triggered.add(103)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 1150) + abs(light - 7500) / 10 + abs(temp - 25) < 500):
        triggered.add(104)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 1150) + abs(light - 8500) / 20 + abs(temp - 25) < 500):
        triggered.add(105)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 15) < 500):
        triggered.add(106)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 400):
        triggered.add(107)

    if (light > 1200 and light < 9800) != (light > 2200 and light < 9800):
        triggered.add(108)
    if (light > 1200 and light < 9800) != (light > 4200 and light < 9800):
        triggered.add(109)
    if (light > 1200 and light < 9800) != (light > 6200 and light < 9800):
        triggered.add(110)
    if (light > 1200 and light < 9800) != (light > 7200 and light < 9800):
        triggered.add(111)
    if (light > 1200 and light < 9800) != (light > 1200 and light < 7800):
        triggered.add(112)
    if (light > 1200 and light < 9800) != (light > 1200 and light < 5800):
        triggered.add(113)
    if (light > 1200 and light < 9800) != (light > 1200 and light < 2800):
        triggered.add(114)
    if (light > 1200 and light < 9800) != (light > 1200 and light < 3800):
        triggered.add(115)
    if (light > 1200 and light < 9800) != (light > 1200 and light < 4800):
        triggered.add(116)

    if (temp > 3 and temp < 38) != (temp > 13 and temp < 38):
        triggered.add(117)
    if (temp > 3 and temp < 38) != (temp > 23 and temp < 38):
        triggered.add(118)
    if (temp > 3 and temp < 38) != (temp > 33 and temp < 38):
        triggered.add(119)
    if (temp > 3 and temp < 38) != (temp > 3 and temp < 8):
        triggered.add(120)
    if (temp > 3 and temp < 38) != (temp > 3 and temp < 18):
        triggered.add(121)
    if (temp > 3 and temp < 38) != (temp > 3 and temp < 28):
        triggered.add(122)
    if (temp > 3 and temp < 38) != (temp > 3 and temp < 30):
        triggered.add(123)

    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 1000 and light > 1500 and temp > 5):
        triggered.add(124)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 950 and light > 1500 and temp > 5):
        triggered.add(125)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 1250 and light > 1500 and temp > 5):
        triggered.add(126)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 or light > 1500 and temp > 5):
        triggered.add(127)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 3500 and temp > 5):
        triggered.add(128)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 5500 and temp > 5):
        triggered.add(129)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 6500 and temp > 5):
        triggered.add(130)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 1500 and temp > 15):
        triggered.add(131)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 1500 and temp > 25):
        triggered.add(132)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 1500 and temp > 35):
        triggered.add(133)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 1500 or temp > 5):
        triggered.add(134)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 2500 and temp > 5):
        triggered.add(135)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 4500 and temp > 5):
        triggered.add(136)

    return triggered


# === SAC Actor ===
class GaussianPolicy(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dim=256):
        super(GaussianPolicy, self).__init__()

        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)

        self.mean = nn.Linear(hidden_dim, action_dim)
        self.log_std = nn.Linear(hidden_dim, action_dim)

        self.action_scale = 8.0
        self.action_bias = 0.0

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))

        mean = self.mean(x)
        log_std = self.log_std(x)
        log_std = torch.clamp(log_std, min=-20, max=2)

        return mean, log_std

    def sample(self, state):
        mean, log_std = self.forward(state)
        std = log_std.exp()
        normal = Normal(mean, std)

        x_t = normal.rsample()
        action = torch.tanh(x_t) * self.action_scale + self.action_bias

        log_prob = normal.log_prob(x_t)
        log_prob -= torch.log(self.action_scale * (1 - torch.tanh(x_t).pow(2)) + 1e-6)
        log_prob = log_prob.sum(1, keepdim=True)

        mean = torch.tanh(mean) * self.action_scale + self.action_bias

        return action, log_prob, mean


# === SAC Critic ===
class QNetwork(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dim=256):
        super(QNetwork, self).__init__()

        self.fc1 = nn.Linear(state_dim + action_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, 1)

        self.fc4 = nn.Linear(state_dim + action_dim, hidden_dim)
        self.fc5 = nn.Linear(hidden_dim, hidden_dim)
        self.fc6 = nn.Linear(hidden_dim, 1)

    def forward(self, state, action):
        xu = torch.cat([state, action], 1)

        x1 = F.relu(self.fc1(xu))
        x1 = F.relu(self.fc2(x1))
        q1 = self.fc3(x1)

        x2 = F.relu(self.fc4(xu))
        x2 = F.relu(self.fc5(x2))
        q2 = self.fc6(x2)

        return q1, q2


# ===  ===
class EnhancedReplayBuffer:
    def __init__(self, capacity=50000):
        self.buffer = deque(maxlen=capacity)
        self.experience_info = deque(maxlen=capacity)

    def push(self, state, action, reward, next_state, done, path_idx, similarity):
        self.buffer.append((state, action, reward, next_state, done))
        self.experience_info.append({'path_idx': path_idx, 'similarity': similarity})

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return None

        batch = random.sample(self.buffer, batch_size)
        state, action, reward, next_state, done = map(np.stack, zip(*batch))

        return (
            torch.FloatTensor(state).to(device),
            torch.FloatTensor(action).to(device),
            torch.FloatTensor(reward).unsqueeze(1).to(device),
            torch.FloatTensor(next_state).to(device),
            torch.FloatTensor(done).unsqueeze(1).to(device)
        )

    def get_top_k_per_path(self, num_paths, k=20):
        path_samples = {i: [] for i in range(num_paths)}

        for idx, info in enumerate(self.experience_info):
            path_idx = info['path_idx']
            similarity = info['similarity']
            path_samples[path_idx].append((idx, similarity, self.buffer[idx]))

        top_k_results = {}
        for path_idx in range(num_paths):
            samples = path_samples[path_idx]
            if len(samples) == 0:
                top_k_results[path_idx] = []
                continue

            samples.sort(key=lambda x: x[1], reverse=True)
            top_k = samples[:k]

            top_k_results[path_idx] = []
            for sample in top_k:
                normalized_state = sample[2][0]
                original_state = denormalize_state(normalized_state)
                original_state_int = np.round(original_state).astype(int)

                # : 
                co2, moisture, temp = original_state_int
                triggered = execute_Tr(co2, moisture, temp)

                top_k_results[path_idx].append({
                    'state': original_state_int,
                    'similarity': sample[1],
                    'triggered': triggered
                })

        return top_k_results

    def __len__(self):
        return len(self.buffer)


# === SAC ===
class SACAgent:
    def __init__(self, state_dim=3, action_dim=3):
        self.state_dim = state_dim
        self.action_dim = action_dim

        # : 
        self.policy = GaussianPolicy(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.policy_optimizer = optim.Adam(self.policy.parameters(), lr=EXPERIMENT_CONFIG['ACTOR_LR'])

        self.critic = QNetwork(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_target = QNetwork(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_target.load_state_dict(self.critic.state_dict())
        self.critic_optimizer = optim.Adam(self.critic.parameters(), lr=EXPERIMENT_CONFIG['CRITIC_LR'])

        self.target_entropy = -action_dim
        self.log_alpha = torch.zeros(1, requires_grad=True, device=device)
        self.alpha_optimizer = optim.Adam([self.log_alpha], lr=EXPERIMENT_CONFIG['ALPHA_LR'])

        self.replay_buffer = EnhancedReplayBuffer()
        self.replay_train_count = 0

    def get_action(self, state, deterministic=False):
        normalized_state = normalize_state(state)
        state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)

        with torch.no_grad():
            if deterministic:
                _, _, action = self.policy.sample(state_tensor)
            else:
                action, _, _ = self.policy.sample(state_tensor)

        action = action.cpu().numpy()[0]
        return action

    def store_experience(self, state, action, reward, next_state, done, path_idx, similarity):
        normalized_state = normalize_state(state)
        normalized_next_state = normalize_state(next_state)

        self.replay_buffer.push(
            normalized_state, action, reward,
            normalized_next_state, done, path_idx, similarity
        )

    def replay_train(self):
        batch = self.replay_buffer.sample(EXPERIMENT_CONFIG['REPLAY_BATCH_SIZE'])

        if batch is None:
            return

        state, action, reward, next_state, done = batch

        with torch.no_grad():
            next_action, next_log_prob, _ = self.policy.sample(next_state)

            q1_next, q2_next = self.critic_target(next_state, next_action)
            q_next = torch.min(q1_next, q2_next)
            target_q = reward + (1 - done) * EXPERIMENT_CONFIG['GAMMA'] * (
                    q_next - self.log_alpha.exp() * next_log_prob)

        q1, q2 = self.critic(state, action)
        critic_loss = F.mse_loss(q1, target_q) + F.mse_loss(q2, target_q)

        self.critic_optimizer.zero_grad()
        critic_loss.backward()
        self.critic_optimizer.step()

        new_action, log_prob, _ = self.policy.sample(state)
        q1_new, q2_new = self.critic(state, new_action)
        q_new = torch.min(q1_new, q2_new)

        policy_loss = (self.log_alpha.exp() * log_prob - q_new).mean()

        self.policy_optimizer.zero_grad()
        policy_loss.backward()
        self.policy_optimizer.step()

        alpha_loss = -(self.log_alpha * (log_prob + self.target_entropy).detach()).mean()

        self.alpha_optimizer.zero_grad()
        alpha_loss.backward()
        self.alpha_optimizer.step()

        for param, target_param in zip(self.critic.parameters(), self.critic_target.parameters()):
            target_param.data.copy_(
                EXPERIMENT_CONFIG['TAU'] * param.data + (1 - EXPERIMENT_CONFIG['TAU']) * target_param.data)

        self.replay_train_count += 1

        if self.replay_train_count % 2 == 0:
            alpha_value = self.log_alpha.exp().item()
            print(f"  ->  (Run {self.replay_train_count}), Alpha={alpha_value:.4f}")


# === Metric ===
def calculate_run_performance(run_idx, sac_results, training_time, total_steps, update_count, agent):
    """ runMetric"""
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    # Similarity
    all_similarities = []

    # Metric
    total_samples = 0
    all_rewards = []
    total_reward = 0

    for path_idx in range(num_paths):
        samples = sac_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']

            total_reward += reward
            all_rewards.append(reward)
            all_similarities.append(similarity)
            total_samples += 1

    # 1. 
    total_reward = total_reward

    # 2. 
    if total_samples > 0:
        average_reward = total_reward / total_samples
    else:
        average_reward = 0

    # 5. (Average Similarity)
    if all_similarities:
        convergence = np.mean(all_similarities)
    else:
        convergence = 0

    # 12. (Similarity)
    if len(all_similarities) > 1:
        environment_adaptability = 1 / (np.std(all_similarities) + 1e-8)
    else:
        environment_adaptability = 0

    # 13. (Average Similarity)
    generalization_ability = convergence

    # 15. (/ seconds)
    if training_time > 0:
        computational_efficiency = total_steps / training_time
    else:
        computational_efficiency = 0

    # 16. 
    if training_time > 0:
        policy_update_frequency = update_count / training_time
    else:
        policy_update_frequency = 0

    # Similarity
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0

    return {
        'Run': run_idx + 1,

        # Metric
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),

        # Similarity
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }


# === Excel ===
# === Excel导出 ===
def export_to_excel(all_sac_results, all_performance_data, target_paths, output_path="SAC_20 run.xlsx"):
    """导出 20 次 run 的 SAC 测试结果到 Excel"""
    print("\n正在导出数据到 Excel...")

    # 初始化收集所有 run 的数据
    all_sac_summary_data = []
    all_sac_detailed_data = []

    # 遍历每次 run 的结果
    for run_idx, (sac_results, performance_data) in enumerate(zip(all_sac_results, all_performance_data)):
        # ===== Sheet1: SACPath 统计 =====
        sac_summary_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = sac_results[path_idx]

            if len(samples) == 0:
                sac_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue  # 结构修复：这里必须有 continue

            # 正常结果计算
            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = 'Yes' if perfect_count > 0 else 'No'

            sac_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })

        all_sac_summary_data.extend(sac_summary_data)

        # ===== Sheet2: SACDetailed Sample Data =====
        sac_detailed_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = sac_results[path_idx]

            # 结构修复：这里必须有 sample 的遍历循环
            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']

                sac_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Sample ID': sample_idx + 1,
                    'CO2': int(state[0]),
                    'Moisture': int(state[1]),
                    'Temperature': int(state[2]),
                    'Similarity': round(similarity, 4),
                    'Perfect Coverage': 'Yes' if abs(similarity - 1.0) < 0.001 else 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path))),
                    'Triggered Rules': ', '.join(map(str, sorted(triggered))),
                    'Intersection Count': len(target_path.intersection(triggered)),
                    'Target Rule Count': len(target_path)
                })

        all_sac_detailed_data.extend(sac_detailed_data)

    # ================= 结构修复：导出过程必须在所有 Run 循环之后 =================
    # 转换为 DataFrame
    sac_summary_df = pd.DataFrame(all_sac_summary_data)
    sac_detailed_df = pd.DataFrame(all_sac_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet1: SACPath 统计
        sac_summary_df.to_excel(writer, sheet_name='SACPath', index=False)

        # Sheet2: SACDetailed Sample Data
        sac_detailed_df.to_excel(writer, sheet_name='SACDetailed Sample Data', index=False)

        # Sheet3: Metric - 性能指标汇总
        selected_columns = [
            'Run',
            'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df_selected = performance_df[selected_columns]
        performance_df_selected.to_excel(writer, sheet_name='Metric', index=False)

        # 获取 workbook 对象进行格式设置
        workbook = writer.book

        # 设置表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        perfect_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # 浅绿色高亮

        # === 设置 Sheet1 样式 ===
        ws1 = writer.sheets['SACPath']
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 高亮完全覆盖的行
        for row_idx in range(2, ws1.max_row + 1):
            if ws1.cell(row_idx, 9).value == 'Yes':  # 第9列是 Perfect Coverage
                for col_idx in range(1, ws1.max_column + 1):
                    ws1.cell(row_idx, col_idx).fill = perfect_fill

        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 18
        ws1.column_dimensions['D'].width = 15
        ws1.column_dimensions['E'].width = 18
        ws1.column_dimensions['F'].width = 15
        ws1.column_dimensions['G'].width = 15
        ws1.column_dimensions['H'].width = 15
        ws1.column_dimensions['I'].width = 18
        ws1.column_dimensions['J'].width = 50

        # === 设置 Sheet2 样式 ===
        ws2 = writer.sheets['SACDetailed Sample Data']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 18
        ws2.column_dimensions['I'].width = 40
        ws2.column_dimensions['J'].width = 40
        ws2.column_dimensions['K'].width = 18
        ws2.column_dimensions['L'].width = 18

        # === 设置 Sheet3 样式 ===
        ws3 = writer.sheets['Metric']
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 调整列宽
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        for col in columns:
            ws3.column_dimensions[col].width = 20

    print(f"文件已成功保存到: {output_path}")
    print(f"  - Sheet1: SACPath 统计共计 {len(all_sac_summary_data)} 条记录")
    print(f"  - Sheet2: SACDetailed Sample Data 共计 {len(all_sac_detailed_data)} 条记录")
    print(f"  - Sheet3: Metric 共计 {len(all_performance_data)} 条记录")

# ===  ===
def train_sac_workflow():
    print("=" * 80)
    print("SAC")
    print("Similarity:  / target paths")
    print("=" * 80)

    agent = SACAgent()
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n: Path {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']}")
    path_samples = {}
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']

    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            # 
            state = np.random.uniform(min_vals, max_vals).astype(np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  Path  {path_idx + 1}/{num_paths}:  {len(samples)} ")

    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size

    print(f"\n: {batch_size},{EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(f": {num_batches} /Path  x {num_paths} Path  = {num_batches * num_paths} ")
    print("-" * 80)

    for batch_idx in range(num_batches):
        print(f"\n {batch_idx + 1}/{num_batches}")

        for path_idx in range(num_paths):
            target_path = target_paths[path_idx]
            batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]

            batch_rewards = []
            batch_similarities = []

            for sample_idx, initial_state in enumerate(batch_samples):
                state = initial_state.copy()
                episode_reward = 0
                final_similarity = 0

                for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                    action = agent.get_action(state, deterministic=False)

                    next_state = state + action
                    next_state = clip_state(next_state)

                    # : 
                    co2, moisture, temp = next_state
                    triggered = execute_Tr(co2, moisture, temp)
                    reward = unified_reward_function(triggered, target_path)
                    similarity = coverage_similarity(triggered, target_path)

                    done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)

                    agent.store_experience(
                        state, action, reward, next_state, done,
                        path_idx, similarity
                    )

                    state = next_state
                    episode_reward += reward
                    final_similarity = similarity
                    total_steps += 1

                batch_rewards.append(episode_reward)
                batch_similarities.append(final_similarity)

            avg_reward = np.mean(batch_rewards)
            avg_similarity = np.mean(batch_similarities)
            print(f"  Path {path_idx + 1}: ={avg_reward:.2f}, Average Similarity={avg_similarity:.4f}")

        print(f"\n  ...")
        agent.replay_train()
        print(f"  : {len(agent.replay_buffer)}")

    training_time = time.time() - start_time

    print("\n" + "=" * 80)
    print(f"SACcompleted! Total elapsed time: {training_time:.2f} seconds, : {total_steps}")
    print(f": {len(agent.replay_buffer)}")
    print(f": {agent.replay_train_count}")
    print("=" * 80)

    print(f"\nPath SimilarityMaximum{EXPERIMENT_CONFIG['TOP_K_SAMPLES']}...")
    top_k_results = agent.replay_buffer.get_top_k_per_path(num_paths, EXPERIMENT_CONFIG['TOP_K_SAMPLES'])

    return agent, top_k_results, training_time, total_steps, agent.replay_train_count


# ===  ===
def main():
    print("\n" + "=" * 80)
    print("SAC - 20 run")
    print(": CO2(800-1500), moisture(10-80), temperature(1-40)")
    print("Metric")
    print("=" * 80)

    all_sac_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    # 20
    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'=' * 80}")
        print(f"Start run  {run_idx + 1}/{EXPERIMENT_CONFIG['NUM_RUNS']}  run")
        print(f"{'=' * 80}")

        # SAC
        sac_agent, sac_results, training_time, total_steps, update_count = train_sac_workflow()

        # Metric
        performance_data = calculate_run_performance(
            run_idx, sac_results, training_time, total_steps, update_count, sac_agent
        )

        # 
        all_sac_results.append(sac_results)
        all_performance_data.append(performance_data)

        print(f"\nRun  {run_idx + 1}  runcompleted!")
        print(f"  Total Reward: {performance_data['Total Reward']}")
        print(f"  Average Reward: {performance_data['Average Reward']}")
        print(f"  Average Similarity: {performance_data['Average Similarity']}")

    # Excel(20 run)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"SAC_20 run_{timestamp}.xlsx"
    export_to_excel(all_sac_results, all_performance_data, target_paths, output_path)

    # 
    print("\n" + "=" * 80)
    print("20 run")
    print("=" * 80)

    # Metric
    total_rewards = [p['Total Reward'] for p in all_performance_data]
    average_rewards = [p['Average Reward'] for p in all_performance_data]
    convergences = [p['Convergence'] for p in all_performance_data]
    environment_adaptabilities = [p['Environment Adaptability'] for p in all_performance_data]
    generalization_abilities = [p['Generalization Ability'] for p in all_performance_data]
    computational_efficiencies = [p['Computational Efficiency'] for p in all_performance_data]
    policy_update_frequencies = [p['Policy Update Frequency'] for p in all_performance_data]
    avg_similarities = [p['Average Similarity'] for p in all_performance_data]

    print(f":")
    print(f"  : {np.mean(total_rewards):.2f}")
    print(f"  Standard deviation: {np.std(total_rewards):.2f}")

    print(f"\n:")
    print(f"  : {np.mean(average_rewards):.4f}")
    print(f"  Standard deviation: {np.std(average_rewards):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(convergences):.4f}")
    print(f"  Standard deviation: {np.std(convergences):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(environment_adaptabilities):.4f}")
    print(f"  Standard deviation: {np.std(environment_adaptabilities):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(generalization_abilities):.4f}")
    print(f"  Standard deviation: {np.std(generalization_abilities):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(computational_efficiencies):.2f}")
    print(f"  Standard deviation: {np.std(computational_efficiencies):.2f}")

    print(f"\n:")
    print(f"  : {np.mean(policy_update_frequencies):.4f}")
    print(f"  Standard deviation: {np.std(policy_update_frequencies):.4f}")

    print(f"\nAverage similarity statistics:")
    print(f"  : {np.mean(avg_similarities):.4f}")
    print(f"  Standard deviation: {np.std(avg_similarities):.4f}")

    print("\n" + "=" * 80)
    print(f" {EXPERIMENT_CONFIG['NUM_RUNS']} completed!")
    print("=" * 80)


if __name__ == "__main__":
    main()