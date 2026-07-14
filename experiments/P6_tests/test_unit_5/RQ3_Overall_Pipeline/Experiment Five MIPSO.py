import os
import random
import numpy as np
import time
import psutil
import math
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

NUM_RUNS = 20  # 默认实验运行次数

# === 新的三维范围设置 ===
LIGHT_MIN = 1
LIGHT_MAX = 40
MOISTURE_MIN = 1000
MOISTURE_MAX = 10000
TEMP_MIN = 800
TEMP_MAX = 1500

BOUNDS = {
    'light': (LIGHT_MIN, LIGHT_MAX),
    'temp': (TEMP_MIN, TEMP_MAX),
    'moisture': (MOISTURE_MIN, MOISTURE_MAX)
}


def clip_state(state):
    """将状态限制在各维度的边界内"""
    return np.array([
        np.clip(state[0], BOUNDS['light'][0], BOUNDS['light'][1]),
        np.clip(state[1], BOUNDS['temp'][0], BOUNDS['temp'][1]),
        np.clip(state[2], BOUNDS['moisture'][0], BOUNDS['moisture'][1])
    ])


def execute_Tr(position):
    """执行目标函数并返回触发的路径"""
    x = int(np.clip(position[0], BOUNDS['light'][0], BOUNDS['light'][1]))
    temp = int(np.clip(position[1], BOUNDS['temp'][0], BOUNDS['temp'][1]))
    z = int(np.clip(position[2], BOUNDS['moisture'][0], BOUNDS['moisture'][1]))
    return category1_multivariable_control(x, temp, z)


# === 目标函数 ===
def category1_multivariable_control(light, co2, temp):
    """高CO2极端条件增强版分支测试函数"""
    actions = []
    triggered = set()

    # Fixed: removed extra space before if
    if (co2 > 1200) != (co2 > 400):
        triggered.add(1)
    if (co2 > 1200) != (co2 > 200):
        triggered.add(2)
    if (co2 > 1200) != (co2 > 100):
        triggered.add(3)
    if (co2 > 1200) != (co2 > 600):
        triggered.add(4)

    if (co2 > 1100 and light > 7000) != (co2 > 100 and light > 7000):
        triggered.add(5)
    if (co2 > 1100 and light > 7000) != (co2 > 500 and light > 7000):
        triggered.add(6)
    if (co2 > 1100 and light > 7000) != (co2 > 800 and light > 7000):
        triggered.add(7)
    if (co2 > 1100 and light > 7000) != (co2 > 1100 and light > 5000):
        triggered.add(8)
    if (co2 > 1100 and light > 7000) != (co2 > 1100 and light > 4000):
        triggered.add(9)
    if (co2 > 1100 and light > 7000) != (co2 > 1100 and light > 3000):
        triggered.add(10)
    if (co2 > 1100 and light > 7000) != (co2 > 1100 and light > 2000):
        triggered.add(11)

    if (co2 > 1250 and temp > 28) != (co2 > 950 and temp > 28):
        triggered.add(12)
    if (co2 > 1250 and temp > 28) != (co2 > 1250 and temp > 18):
        triggered.add(13)
    if (co2 > 1250 and temp > 28) != (co2 > 1000 and temp > 28):
        triggered.add(14)
    if (co2 > 1250 and temp > 28) != (co2 > 900 and temp > 28):
        triggered.add(15)
    if (co2 > 1250 and temp > 28) != (co2 > 1250 or temp > 28):
        triggered.add(16)
    if (co2 > 1250 and temp > 28) != (co2 > 1250 and temp <= 28):
        triggered.add(17)
    if (co2 > 1250 and temp > 28) != (co2 < 1250 and temp > 28):
        triggered.add(18)
    if (co2 > 1250 and temp > 28) != (light > 1250 and temp > 28):
        triggered.add(19)

    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1350 and light > 6000 and temp > 22):
        triggered.add(20)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 < 1150 and light > 6000 and temp > 22):
        triggered.add(21)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 7000 and temp > 22):
        triggered.add(22)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 8000 and temp > 22):
        triggered.add(23)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 4000 and temp > 22):
        triggered.add(24)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 6000 and temp > 32):
        triggered.add(25)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 6000 and temp > 12):
        triggered.add(26)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light < 6000 and temp > 22):
        triggered.add(27)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 6000 and temp < 22):
        triggered.add(28)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 or light > 6000 and temp > 22):
        triggered.add(29)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 6000 or temp > 22):
        triggered.add(30)

    if (co2 < 1050 and light > 6000) != (co2 < 1250 and light > 6000):
        triggered.add(31)
    if (co2 < 1050 and light > 6000) != (co2 < 1350 and light > 6000):
        triggered.add(32)
    if (co2 < 1050 and light > 6000) != (co2 < 1450 and light > 6000):
        triggered.add(33)
    if (co2 < 1050 and light > 6000) != (co2 > 1050 and light > 6000):
        triggered.add(34)
    if (co2 < 1050 and light > 6000) != (co2 < 1050 or light > 6000):
        triggered.add(35)
    if (co2 < 1050 and light > 6000) != (co2 < 1050 and light < 6000):
        triggered.add(36)
    if (co2 < 1050 and light > 6000) != (co2 < 1050 and light > 4000):
        triggered.add(37)
    if (co2 < 1050 and light > 6000) != (co2 < 1050 and light > 8000):
        triggered.add(38)

    if (temp > 28 and light > 6000) != (temp > 18 and light > 6000):
        triggered.add(39)
    if (temp > 28 and light > 6000) != (temp > 24 and light > 6000):
        triggered.add(40)
    if (temp > 28 and light > 6000) != (temp > 20 and light > 6000):
        triggered.add(41)
    if (temp > 28 and light > 6000) != (temp > 28 or light > 6000):
        triggered.add(42)
    if (temp > 28 and light > 6000) != (temp > 28 and light > 4000):
        triggered.add(43)
    if (temp > 28 and light > 6000) != (temp > 28 and light > 2000):
        triggered.add(44)
    if (temp > 28 and light > 6000) != (temp > 28 and light > 8000):
        triggered.add(45)

    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1300 and light > 6000 and temp > 22):
        triggered.add(46)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 600 and light > 6000 and temp > 22):
        triggered.add(47)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 300 and light > 6000 and temp > 22):
        triggered.add(48)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 < 1100 and light > 6000 and temp > 22):
        triggered.add(49)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1100 or light > 6000 and temp > 22):
        triggered.add(50)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1100 and light > 4000 and temp > 22):
        triggered.add(51)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1100 and light > 8000 and temp > 22):
        triggered.add(52)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1100 and light > 6000 or temp > 22):
        triggered.add(53)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1100 and light > 6000 and temp > 12):
        triggered.add(54)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1100 and light > 6000 and temp > 32):
        triggered.add(55)

    if (co2 < 1050 and light > 5000) != (co2 < 1250 and light > 5000):
        triggered.add(56)
    if (co2 < 1050 and light > 5000) != (co2 < 950 and light > 5000):
        triggered.add(57)
    if (co2 < 1050 and light > 5000) != (co2 < 1050 or light > 5000):
        triggered.add(58)
    if (co2 < 1050 and light > 5000) != (co2 < 1050 and light > 6000):
        triggered.add(59)
    if (co2 < 1050 and light > 5000) != (co2 < 1050 and light > 8000):
        triggered.add(60)
    if (co2 < 1050 and light > 5000) != (co2 < 1050 and light > 3000):
        triggered.add(61)

    if (3000 <= light <= 8000) != (2000 <= light <= 8000):
        triggered.add(62)
    if (3000 <= light <= 8000) != (1000 <= light <= 8000):
        triggered.add(63)
    if (3000 <= light <= 8000) != (4000 <= light <= 8000):
        triggered.add(64)
    if (3000 <= light <= 8000) != (6000 <= light <= 8000):
        triggered.add(65)

    if (co2 > 1180 and light > 5500) != (co2 > 1280 and light > 5500):
        triggered.add(66)
    if (co2 > 1180 and light > 5500) != (co2 > 1380 and light > 5500):
        triggered.add(67)
    if (co2 > 1180 and light > 5500) != (co2 < 1180 and light > 5500):
        triggered.add(68)
    if (co2 > 1180 and light > 5500) != (co2 > 1180 or light > 5500):
        triggered.add(69)
    if (co2 > 1180 and light > 5500) != (co2 > 1180 and light > 6600):
        triggered.add(70)
    if (co2 > 1180 and light > 5500) != (co2 > 1180 and light > 7700):
        triggered.add(71)
    if (co2 > 1180 and light > 5500) != (co2 > 1180 and light > 8800):
        triggered.add(72)

    if (light > 3500 and light < 8500) != (light > 5500 and light < 8500):
        triggered.add(73)
    if (light > 3500 and light < 8500) != (light > 6500 and light < 8500):
        triggered.add(74)
    if (light > 3500 and light < 8500) != (light > 7500 and light < 8500):
        triggered.add(75)
    if (light > 3500 and light < 8500) != (light < 3500 and light < 8500):
        triggered.add(76)
    if (light > 3500 and light < 8500) != (light > 3500 and light > 8500):
        triggered.add(77)
    if (light > 3500 and light < 8500) != (light > 3500 and light < 7500):
        triggered.add(78)
    if (light > 3500 and light < 8500) != (light > 3500 and light < 5500):
        triggered.add(79)
    if (light > 3500 and light < 8500) != (light > 3500 and light < 4500):
        triggered.add(80)

    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 900) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000):
        triggered.add(81)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1250) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000):
        triggered.add(82)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 3 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000):
        triggered.add(83)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 4500) ** 2 / 100 + (temp - 25) ** 2 < 10000):
        triggered.add(84)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 3500) ** 2 / 100 + (temp - 25) ** 2 < 10000):
        triggered.add(85)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 5500) ** 3 / 100 + (temp - 25) ** 2 < 10000):
        triggered.add(86)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 5500) ** 2 / 200 + (temp - 25) ** 2 < 10000):
        triggered.add(87)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 15) ** 2 < 10000):
        triggered.add(88)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 35) ** 2 < 10000):
        triggered.add(89)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 3 < 10000):
        triggered.add(90)

    if (light / temp > 50 and light / temp < 300) != (light / temp > 60 and light / temp < 300):
        triggered.add(91)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 70 and light / temp < 300):
        triggered.add(92)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 80 and light / temp < 300):
        triggered.add(93)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 90 and light / temp < 300):
        triggered.add(94)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 50 or light / temp < 300):
        triggered.add(95)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 50 and light / temp < 200):
        triggered.add(96)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 50 and light / temp < 100):
        triggered.add(97)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 50 and light / temp < 350):
        triggered.add(98)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 50 and light / temp < 400):
        triggered.add(99)

    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 1050) + abs(light - 5500) / 10 + abs(temp - 25) < 500):
        triggered.add(100)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 950) + abs(light - 5500) / 10 + abs(temp - 25) < 500):
        triggered.add(101)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 850) + abs(light - 5500) / 10 + abs(temp - 25) < 500):
        triggered.add(102)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 1150) + abs(light - 6500) / 10 + abs(temp - 25) < 500):
        triggered.add(103)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 1150) + abs(light - 7500) / 10 + abs(temp - 25) < 500):
        triggered.add(104)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 1150) + abs(light - 8500) / 20 + abs(temp - 25) < 500):
        triggered.add(105)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 15) < 500):
        triggered.add(106)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 400):
        triggered.add(107)

    if (light > 1200 and light < 9800) != (light > 2200 and light < 9800):
        triggered.add(108)
    if (light > 1200 and light < 9800) != (light > 4200 and light < 9800):
        triggered.add(109)
    if (light > 1200 and light < 9800) != (light > 6200 and light < 9800):
        triggered.add(110)
    if (light > 1200 and light < 9800) != (light > 7200 and light < 9800):
        triggered.add(111)
    if (light > 1200 and light < 9800) != (light > 1200 and light < 7800):
        triggered.add(112)
    if (light > 1200 and light < 9800) != (light > 1200 and light < 5800):
        triggered.add(113)
    if (light > 1200 and light < 9800) != (light > 1200 and light < 2800):
        triggered.add(114)
    if (light > 1200 and light < 9800) != (light > 1200 and light < 3800):
        triggered.add(115)
    if (light > 1200 and light < 9800) != (light > 1200 and light < 4800):
        triggered.add(116)

    if (temp > 3 and temp < 38) != (temp > 13 and temp < 38):
        triggered.add(117)
    if (temp > 3 and temp < 38) != (temp > 23 and temp < 38):
        triggered.add(118)
    if (temp > 3 and temp < 38) != (temp > 33 and temp < 38):
        triggered.add(119)
    if (temp > 3 and temp < 38) != (temp > 3 and temp < 8):
        triggered.add(120)
    if (temp > 3 and temp < 38) != (temp > 3 and temp < 18):
        triggered.add(121)
    if (temp > 3 and temp < 38) != (temp > 3 and temp < 28):
        triggered.add(122)
    if (temp > 3 and temp < 38) != (temp > 3 and temp < 30):
        triggered.add(123)

    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 1000 and light > 1500 and temp > 5):
        triggered.add(124)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 950 and light > 1500 and temp > 5):
        triggered.add(125)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 1250 and light > 1500 and temp > 5):
        triggered.add(126)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 or light > 1500 and temp > 5):
        triggered.add(127)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 3500 and temp > 5):
        triggered.add(128)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 5500 and temp > 5):
        triggered.add(129)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 6500 and temp > 5):
        triggered.add(130)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 1500 and temp > 15):
        triggered.add(131)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 1500 and temp > 25):
        triggered.add(132)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 1500 and temp > 35):
        triggered.add(133)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 1500 or temp > 5):
        triggered.add(134)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 2500 and temp > 5):
        triggered.add(135)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 4500 and temp > 5):
        triggered.add(136)

    return triggered


# 目标路径定义
targetPaths = [
    [1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20, 21, 22, 23, 25, 27, 28, 31, 32, 33, 34, 35, 45, 46, 49,
     52, 55, 56, 58, 66, 67, 68, 70, 71, 72, 74, 75, 76, 77, 79, 80, 81, 82, 83, 84, 85, 86, 96, 97, 111, 113, 114,
     115, 116, 119, 120, 121, 122, 123, 126, 130, 133],
    [1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20, 21, 22, 23, 25, 27, 28, 31, 32, 33, 34, 35, 45, 46, 49, 52,
     55, 56, 58, 66, 67, 68, 70, 71, 72, 74, 75, 76, 77, 79, 80, 81, 83, 84, 85, 86, 97, 110, 111, 113, 114, 115, 116,
     119, 120, 121, 122, 123, 126, 130, 133],
    [1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20, 21, 22, 23, 25, 27, 28, 31, 32, 33, 34, 35, 45, 46, 49, 52,
     55, 56, 58, 68, 69, 74, 75, 76, 77, 79, 80, 81, 82, 84, 85, 86, 96, 97, 110, 111, 113, 114, 115, 116, 119, 120,
     121, 122, 123, 126, 130, 133],
    [8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20, 21, 22, 23, 27, 28, 31, 32, 33, 34, 35, 45, 46, 49, 52, 56, 58, 66, 67,
     68, 70, 71, 72, 74, 75, 76, 77, 79, 80, 81, 83, 84, 85, 86, 90, 97, 110, 111, 113, 114, 115, 116, 120, 121, 122,
     123, 126, 130],
    [1, 2, 3, 4, 8, 9, 10, 11, 26, 28, 29, 30, 31, 32, 33, 34, 35, 39, 41, 42, 50, 53, 54, 56, 58, 66, 67, 68, 70, 71,
     72, 74, 75, 76, 77, 79, 80, 81, 83, 84, 85, 86, 96, 97, 110, 111, 113, 114, 115, 116, 118, 119, 120, 121, 126,
     130, 132, 133],
    [1, 2, 3, 4, 8, 9, 10, 11, 20, 21, 22, 23, 25, 27, 28, 31, 32, 33, 34, 35, 39, 40, 41, 42, 46, 49, 52, 55, 56, 58,
     68, 69, 74, 75, 76, 77, 79, 80, 81, 82, 84, 85, 86, 96, 97, 110, 111, 113, 114, 115, 116, 119, 120, 121, 126, 130,
     133],
    [1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 24, 27, 29, 30, 42, 43, 44, 50, 51, 53, 56, 58, 65, 66, 67, 68,
     70, 71, 72, 74, 75, 76, 77, 79, 80, 81, 83, 84, 85, 86, 97, 110, 111, 113, 114, 115, 116, 120, 121, 122, 123, 126,
     130],
    [1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20, 21, 22, 23, 25, 27, 28, 31, 32, 33, 34, 35, 45, 46, 49, 52,
     55, 56, 58, 66, 67, 68, 71, 72, 75, 76, 77, 79, 80, 87, 96, 97, 111, 113, 114, 115, 116, 119, 120, 121, 122, 126,
     133],
    [1, 2, 3, 4, 8, 9, 10, 11, 26, 28, 29, 30, 31, 32, 33, 34, 35, 39, 42, 50, 53, 54, 56, 58, 68, 69, 74, 75, 76, 77,
     79, 80, 81, 82, 83, 84, 85, 86, 95, 98, 99, 110, 111, 113, 114, 115, 116, 118, 119, 120, 121, 126, 130, 132, 133],
    [1, 2, 3, 4, 8, 9, 10, 11, 26, 28, 29, 30, 31, 32, 33, 34, 35, 42, 50, 53, 54, 56, 58, 66, 67, 68, 70, 71, 72, 74,
     75, 76, 77, 79, 80, 81, 82, 83, 84, 85, 86, 95, 111, 113, 114, 115, 116, 118, 119, 120, 126, 130, 131, 132, 133],
    [1, 2, 3, 4, 12, 14, 15, 16, 18, 19, 21, 29, 30, 31, 32, 33, 34, 35, 45, 47, 48, 49, 50, 53, 56, 58, 68, 69, 74,
     75, 76, 77, 79, 80, 81, 82, 84, 85, 86, 97, 110, 111, 113, 114, 115, 116, 119, 120, 121, 122, 123, 126, 130, 133],
    [8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20, 21, 22, 23, 25, 27, 28, 31, 32, 33, 34, 35, 45, 46, 49, 52, 55, 56, 58,
     66, 67, 68, 70, 71, 72, 75, 76, 77, 79, 80, 96, 97, 102, 111, 113, 114, 115, 116, 119, 120, 121, 122, 126, 133],
    [1, 2, 3, 4, 8, 9, 10, 11, 28, 29, 30, 31, 32, 33, 34, 35, 42, 50, 53, 56, 58, 66, 67, 68, 70, 71, 72, 74, 75, 76,
     77, 79, 80, 81, 83, 84, 85, 86, 95, 110, 111, 113, 114, 115, 116, 117, 118, 119, 120, 126, 130, 131, 132, 133],
    [1, 2, 3, 4, 5, 6, 7, 16, 18, 19, 21, 29, 30, 34, 36, 38, 45, 47, 48, 49, 50, 53, 60, 68, 69, 76, 77, 78, 79, 80,
     83, 96, 97, 100, 101, 102, 103, 104, 105, 112, 113, 114, 115, 116, 119, 120, 121, 122, 123, 124, 125, 126, 133],
    [1, 2, 3, 4, 12, 14, 15, 16, 18, 19, 21, 29, 30, 31, 32, 33, 34, 35, 45, 47, 48, 49, 50, 53, 56, 58, 68, 69, 74,
     75, 76, 77, 79, 80, 81, 82, 84, 85, 86, 88, 90, 97, 111, 113, 114, 115, 116, 120, 121, 122, 123, 126, 130, 133],
    [8, 9, 10, 11, 13, 16, 17, 20, 21, 22, 23, 25, 27, 28, 32, 33, 34, 35, 39, 41, 42, 46, 49, 52, 55, 58, 67, 68, 70,
     71, 72, 74, 75, 76, 77, 79, 80, 82, 96, 97, 102, 110, 111, 113, 114, 115, 116, 119, 120, 121, 130, 132, 133],
    [8, 9, 10, 11, 26, 28, 29, 30, 31, 32, 33, 34, 35, 42, 50, 53, 54, 56, 58, 66, 67, 68, 70, 71, 72, 74, 75, 76, 77,
     79, 80, 81, 83, 84, 85, 86, 89, 95, 99, 111, 113, 114, 115, 116, 118, 119, 120, 121, 126, 130, 132, 133],
    [8, 9, 10, 11, 13, 16, 17, 21, 22, 23, 25, 27, 28, 34, 35, 39, 41, 42, 49, 52, 55, 58, 68, 70, 71, 72, 74, 75, 76,
     77, 79, 80, 96, 97, 100, 101, 102, 107, 111, 113, 114, 115, 116, 118, 119, 120, 121, 130, 132, 133],
    [1, 2, 3, 4, 5, 6, 7, 16, 18, 19, 21, 29, 30, 34, 36, 38, 45, 47, 48, 49, 50, 53, 60, 68, 69, 75, 76, 77, 79, 80,
     83, 96, 97, 100, 101, 102, 103, 104, 105, 113, 114, 115, 116, 119, 120, 121, 122, 123, 127, 134],
    [1, 2, 3, 4, 5, 6, 7, 15, 16, 18, 19, 21, 29, 30, 34, 36, 38, 45, 47, 48, 49, 50, 53, 60, 68, 69, 76, 77, 78, 79,
     80, 83, 96, 97, 106, 107, 112, 113, 114, 115, 116, 119, 120, 121, 122, 123, 124, 125, 126, 133],
    [1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 24, 27, 29, 30, 42, 43, 44, 50, 51, 53, 56, 58, 65, 73, 74, 75,
     76, 77, 80, 81, 85, 97, 110, 111, 114, 115, 116, 119, 120, 121, 122, 123, 126, 129, 130, 133],
    [1, 2, 3, 4, 12, 15, 16, 18, 19, 21, 29, 30, 34, 36, 38, 45, 47, 48, 49, 50, 53, 57, 60, 68, 69, 74, 75, 76, 77,
     79, 80, 81, 83, 96, 97, 110, 111, 113, 114, 115, 116, 119, 120, 121, 122, 124, 126, 130, 133],
    [1, 2, 3, 4, 9, 10, 11, 12, 14, 15, 16, 18, 19, 24, 27, 29, 30, 42, 43, 44, 50, 51, 53, 65, 69, 73, 74, 75, 76,
     77, 84, 85, 86, 87, 97, 110, 111, 114, 115, 119, 120, 121, 122, 123, 126, 129, 130, 133, 136],
    [1, 2, 3, 4, 9, 10, 11, 12, 14, 15, 16, 18, 19, 24, 27, 29, 30, 42, 43, 44, 50, 51, 53, 65, 73, 74, 75, 76, 77,
     84, 85, 86, 97, 109, 110, 111, 114, 115, 119, 120, 121, 122, 123, 126, 129, 130, 133, 136],
    [1, 2, 3, 4, 12, 14, 15, 16, 18, 19, 27, 29, 30, 42, 50, 53, 63, 69, 76, 86, 91, 92, 93, 94, 100, 101, 102, 103,
     104, 107, 108, 109, 110, 111, 119, 120, 121, 122, 123, 126, 128, 129, 130, 133, 135, 136],
    [1, 2, 3, 4, 16, 18, 19, 30, 35, 36, 37, 42, 43, 44, 53, 58, 61, 65, 73, 74, 75, 76, 77, 83, 86, 97, 103, 104,
     105, 107, 110, 111, 114, 115, 119, 120, 121, 122, 123, 124, 125, 126, 129, 130, 133, 136],
    [1, 2, 3, 4, 12, 15, 16, 18, 19, 30, 35, 36, 42, 44, 53, 58, 62, 63, 76, 83, 86, 92, 93, 94, 100, 101, 102, 105,
     108, 109, 110, 111, 119, 120, 121, 122, 123, 124, 126, 128, 129, 130, 133, 135, 136],
    [1, 2, 3, 4, 16, 18, 19, 30, 35, 36, 42, 44, 53, 58, 61, 64, 65, 73, 74, 75, 76, 77, 83, 86, 97, 103, 104, 105,
     107, 109, 110, 111, 114, 115, 120, 121, 122, 123, 124, 125, 126, 129, 130, 133, 136],
    [1, 2, 3, 4, 15, 16, 18, 19, 30, 35, 36, 37, 42, 43, 44, 53, 59, 60, 65, 68, 69, 74, 75, 76, 77, 79, 80, 81, 83,
     96, 97, 110, 111, 113, 114, 115, 116, 119, 120, 121, 122, 124, 125, 126, 130, 133]
]


# === 增强版指标收集器 ===
class MetricsCollector:
    def __init__(self):
        self.pso_start_time = None
        self.pso_end_time = None
        self.perfect_solutions_count = 0
        self.total_paths_count = 0
        self.final_fitness_scores = []
        self.pso_convergence_iterations = []
        self.pso_reset_counts = []
        self.path_execution_times = []

    def start_pso_phase(self):
        self.pso_start_time = time.time()

    def end_pso_phase(self):
        self.pso_end_time = time.time()

    def record_pso_result(self, fitness, is_perfect_match, convergence_iter=None, path_id=None, method='MI-PSO',
                          reset_count=0, execution_time=0):
        self.final_fitness_scores.append(fitness)
        self.total_paths_count += 1
        self.pso_reset_counts.append(reset_count)
        self.path_execution_times.append(execution_time)

        if is_perfect_match:
            self.perfect_solutions_count += 1
        if convergence_iter is not None:
            self.pso_convergence_iterations.append(convergence_iter)


# === 粒子类 ===
class Particle:
    def __init__(self):
        self.position = np.array([
            np.random.uniform(BOUNDS['light'][0], BOUNDS['light'][1]),
            np.random.uniform(BOUNDS['temp'][0], BOUNDS['temp'][1]),
            np.random.uniform(BOUNDS['moisture'][0], BOUNDS['moisture'][1])
        ])
        self.velocity = np.array([
            np.random.uniform(-5, 5),
            np.random.uniform(-3, 3),
            np.random.uniform(-5, 5)
        ])
        self.best_position = self.position.copy()
        self.best_fitness = 0
        self.fitness = 0


# === 单独的 MI-PSO 优化器类（已完全修复随机性重评估Bug） ===
class MIPSO:
    def __init__(self, target_path, swarm_size=20, max_iterations=3000, g2=10, th_cv=1.2):
        self.target_path = target_path
        self.swarm_size = swarm_size
        self.max_iterations = max_iterations
        self.g2 = g2
        self.th_cv = th_cv

        # 完全随机初始化
        self.particles = [Particle() for _ in range(swarm_size)]
        self.global_best_position = None
        self.global_best_fitness = 0
        self.global_best_triggered = set()  # 核心修复：记录最高分对应的真实触发路径
        self.reset_count = 0

        # 初始化适应度评估
        for particle in self.particles:
            fit, trig = self.evaluate(particle.position)
            particle.fitness = fit
            particle.best_fitness = fit
            particle.best_position = particle.position.copy()

            if fit > self.global_best_fitness:
                self.global_best_fitness = fit
                self.global_best_position = particle.position.copy()
                self.global_best_triggered = trig

    def evaluate(self, position):
        """同时返回适应度和实际触发的路径"""
        try:
            triggered = execute_Tr(position)
            if self.target_path.issubset(triggered):
                return 1.0, triggered
            intersection = len(triggered & self.target_path)
            union = len(triggered | self.target_path)
            return (intersection / union if union > 0 else 0.0), triggered
        except:
            return 0.0, set()

    def calculate_cv(self, scores):
        """实现公式 25：综合变异系数"""
        N = self.swarm_size
        mean_fit = np.mean(scores)
        if mean_fit == 0: return float('inf')

        var_term = np.sum((scores - mean_fit) ** 2) / N
        adj_diff = sum(abs(scores[2 * i + 1] - scores[2 * i]) for i in range(N // 2))
        adj_term = ((2 / N) * adj_diff) ** 2

        numerator = math.sqrt(var_term + adj_term)
        denominator = 2 * np.sum(scores)

        if denominator == 0: return float('inf')
        return numerator / denominator

    def flip_mutation(self, position, rho=0.1):
        """实现公式 27：变异反转操作"""
        new_pos = np.copy(position)
        new_pos[0] = BOUNDS['light'][0] + BOUNDS['light'][1] - position[0] + rho
        new_pos[1] = BOUNDS['temp'][0] + BOUNDS['temp'][1] - position[1] + rho
        new_pos[2] = BOUNDS['moisture'][0] + BOUNDS['moisture'][1] - position[2] + rho
        return clip_state(new_pos)

    def optimize(self):
        w = 0.7
        c1 = 1.5
        c2 = 1.5

        max_velocity = np.array([
            (BOUNDS['light'][1] - BOUNDS['light'][0]) * 0.2,
            (BOUNDS['temp'][1] - BOUNDS['temp'][0]) * 0.2,
            (BOUNDS['moisture'][1] - BOUNDS['moisture'][0]) * 0.2
        ])

        c = 0
        early_stop = False
        converged_at = self.max_iterations

        for iteration in range(self.max_iterations):
            scores = np.array([p.fitness for p in self.particles])
            c += 1

            # MI-PSO 特有的停滞检测与变异 (对应 Algorithm 4 Line 20-23)
            if c == self.g2:
                cv = self.calculate_cv(scores)
                if cv <= self.th_cv:
                    self.reset_count += 1
                    l = int(round(self.swarm_size * (1 - cv)))
                    l = min(max(l, 1), self.swarm_size)

                    selected_indices = np.random.choice(self.swarm_size, l, replace=False)
                    for idx in selected_indices:
                        new_pos = self.flip_mutation(self.particles[idx].position)
                        new_fit, new_trig = self.evaluate(new_pos)

                        if new_fit > self.particles[idx].fitness:
                            self.particles[idx].position = new_pos
                            self.particles[idx].fitness = new_fit
                            if new_fit > self.particles[idx].best_fitness:
                                self.particles[idx].best_fitness = new_fit
                                self.particles[idx].best_position = new_pos.copy()
                            if new_fit > self.global_best_fitness:
                                self.global_best_fitness = new_fit
                                self.global_best_position = new_pos.copy()
                                self.global_best_triggered = new_trig  # 更新最优路径
                c = 0

                # 标准 PSO 速度与位置更新
            for particle in self.particles:
                r1, r2 = np.random.random(3), np.random.random(3)
                particle.velocity = (w * particle.velocity +
                                     c1 * r1 * (particle.best_position - particle.position) +
                                     c2 * r2 * (self.global_best_position - particle.position))
                particle.velocity = np.clip(particle.velocity, -max_velocity, max_velocity)

                particle.position += particle.velocity
                particle.position = clip_state(particle.position)

                fit, trig = self.evaluate(particle.position)
                particle.fitness = fit

                if fit > particle.best_fitness:
                    particle.best_fitness = fit
                    particle.best_position = particle.position.copy()
                if fit > self.global_best_fitness:
                    self.global_best_fitness = fit
                    self.global_best_position = particle.position.copy()
                    self.global_best_triggered = trig  # 更新最优路径

            if self.global_best_fitness >= 1.0:
                converged_at = iteration + 1
                early_stop = True
                break

        return self.global_best_position, self.global_best_fitness, self.global_best_triggered, converged_at, early_stop, self.reset_count


# === Excel导出函数 ===
def export_multiple_runs_to_excel(all_run_results, all_run_metrics, num_runs, filename=None):
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Isolated_MIPSO_{num_runs}Runs_{timestamp}.xlsx"

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    border = Border(
        left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 工作表1: 运行汇总
    ws1 = wb.active
    ws1.title = "运行汇总"
    ws1.sheet_view.showGridLines = False

    headers = ["运行次数", "成功率", "成功数量", "平均适应度", "平均迭代次数", "运行时间(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, border, center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, (results, run_metrics) in enumerate(zip(all_run_results, all_run_metrics), start=1):
        success_count = sum(1 for r in results if r['perfect_match'])
        success_rate = (success_count / len(targetPaths)) * 100
        avg_fitness = np.mean([r['fitness'] for r in results])

        iterations_list = [r['convergence_iteration'] if r.get('convergence_iteration') is not None else 10000 for r in
                           results]
        avg_iterations = np.mean(iterations_list)
        total_time = run_metrics.pso_end_time - run_metrics.pso_start_time if run_metrics.pso_end_time else 0

        row_data = [f"运行 {run_idx}", f"{success_rate:.1f}%", f"{success_count}/{len(targetPaths)}",
                    f"{avg_fitness:.4f}", f"{avg_iterations:.1f}", f"{total_time:.2f}"]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border, cell.alignment = border, center_align
            if run_idx % 2 == 0: cell.fill = alternate_fill
            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    # 工作表2: 路径统计
    ws2 = wb.create_sheet(title="路径统计")
    ws2.sheet_view.showGridLines = False
    headers2 = ["路径编号", "成功次数", "成功率", "平均适应度", "平均迭代次数", "最小迭代", "最大迭代"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, border, center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(targetPaths)):
        success_count = sum(1 for results in all_run_results if results[path_idx]['perfect_match'])
        success_rate = (success_count / num_runs) * 100
        avg_fitness = np.mean([results[path_idx]['fitness'] for results in all_run_results])

        iterations_list = [results[path_idx]['convergence_iteration'] if results[path_idx].get(
            'convergence_iteration') is not None else 10000 for results in all_run_results]
        avg_iterations = np.mean(iterations_list)

        row_data = [f"路径 {path_idx + 1}", f"{success_count}/{num_runs}", f"{success_rate:.1f}%",
                    f"{avg_fitness:.4f}", f"{avg_iterations:.1f}", np.min(iterations_list), np.max(iterations_list)]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border, cell.alignment = border, center_align
            if (path_idx + 1) % 2 == 0: cell.fill = alternate_fill
            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    # 工作表3: 最佳粒子详情
    ws3 = wb.create_sheet(title="最佳粒子详情")
    ws3.sheet_view.showGridLines = False
    headers3 = ["路径", "运行", "最佳粒子(light,temp,moisture)", "适应度", "迭代次数", "求解方法", "生成路径"]
    col_widths3 = [10, 10, 25, 12, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, border, center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(targetPaths)):
        for run_idx, results in enumerate(all_run_results, start=1):
            result = results[path_idx]
            particle_str = f"({int(result['best_position'][0])}, {int(result['best_position'][1])}, {int(result['best_position'][2])})"
            path_str = str(sorted(list(result['triggered'])))

            convergence_iter = result['convergence_iteration'] if result.get(
                'convergence_iteration') is not None else 10000

            row_data = [f"路径{path_idx + 1}", f"运行{run_idx}", particle_str, f"{result['fitness']:.4f}",
                        convergence_iter if convergence_iter < 10000 else "-", "MI-PSO", path_str]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = left_align if col == 7 else center_align

                if result['fitness'] == 1.0:
                    cell.fill = success_fill
                elif result['fitness'] < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill
            row_idx += 1

    # 保存
    output_dir = os.path.join(os.getcwd(), "results")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"\n✓ 独立 MI-PSO 消融实验结果已导出到: {filepath}")
    return filepath


def run_single_mipso_experiment(run_num):
    print(f"\n{'=' * 50}")
    print(f"开始第 {run_num} 次运行 (独立 MI-PSO)")
    print(f"{'=' * 50}")

    run_metrics = MetricsCollector()
    run_metrics.start_pso_phase()

    pso_results = []

    for i, target_path in enumerate(targetPaths):
        path_start_time = time.time()

        mipso = MIPSO(target_path, swarm_size=20, max_iterations=3000)
        # 直接接收并使用最佳寻优过程中真实触发的 best_trig，坚决不重新 evaluate
        best_pos, best_fit, best_trig, converged_at, early_stop, reset_count = mipso.optimize()

        path_execution_time = time.time() - path_start_time

        # 使用真实的最高分数判断是否完美
        is_perfect = (best_fit >= 1.0)

        pso_results.append({
            'target_path': target_path,
            'best_position': best_pos,
            'fitness': best_fit,
            'triggered': best_trig,
            'perfect_match': is_perfect,
            'method': 'MI-PSO',
            'convergence_iteration': converged_at,
            'early_stopped': early_stop,
            'reset_count': reset_count
        })

        run_metrics.record_pso_result(
            fitness=best_fit,
            is_perfect_match=is_perfect,
            convergence_iter=converged_at if early_stop else None,
            path_id=i + 1,
            method='MI-PSO',
            reset_count=reset_count,
            execution_time=path_execution_time
        )

        status = "✓完美(MI-PSO)" if is_perfect else f"○部分({best_fit:.3f})"
        print(f"  路径{i + 1}: {status} | 变异触发次数: {reset_count} | 耗时 {path_execution_time:.2f}s")

    run_metrics.end_pso_phase()

    success_count = sum(1 for r in pso_results if r['perfect_match'])
    success_rate = (success_count / len(targetPaths)) * 100
    pso_time = run_metrics.pso_end_time - run_metrics.pso_start_time

    print(f"\n第{run_num}次运行完成: 成功率 {success_rate:.1f}% | 耗时 {pso_time:.2f}秒")
    return pso_results, run_metrics


def run_multiple_mipso_experiments(num_runs):
    all_run_results = []
    all_run_metrics = []
    total_start = time.time()

    for run_num in range(1, num_runs + 1):
        results, metrics = run_single_mipso_experiment(run_num)
        all_run_results.append(results)
        all_run_metrics.append(metrics)

    total_time = time.time() - total_start
    print(f"\n{'=' * 50}\n全部{num_runs}次独立 MI-PSO 运行完成! 总耗时: {total_time:.2f}秒\n{'=' * 50}")

    return all_run_results, all_run_metrics


if __name__ == "__main__":
    if len(sys.argv) > 1:
        try:
            NUM_RUNS = int(sys.argv[1])
        except ValueError:
            pass

    print("=" * 70)
    print(" 独立 MI-PSO 消融实验 (无 DQN 介入，完美修复随机性 Bug)")
    print("=" * 70)

    all_run_results, all_run_metrics = run_multiple_mipso_experiments(NUM_RUNS)
    export_multiple_runs_to_excel(all_run_results, all_run_metrics, NUM_RUNS)