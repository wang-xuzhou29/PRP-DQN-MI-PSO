import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

def execute_validation_rules(light, co2, temp):
    """高CO2极端条件增强版分支测试函数"""
    actions = []
    triggered = set()

    # Fixed: removed extra space before if
    if (co2 > 1200) != (co2 > 400):
        triggered.add(1)
    if (co2 > 1200) != (co2 > 200):
        triggered.add(2)
    if (co2 > 1200) != (co2 > 100):
        triggered.add(3)
    if (co2 > 1200) != (co2 > 600):
        triggered.add(4)

    if (co2 > 1100 and light > 7000) != (co2 > 100 and light > 7000):
        triggered.add(5)
    if (co2 > 1100 and light > 7000) != (co2 > 500 and light > 7000):
        triggered.add(6)
    if (co2 > 1100 and light > 7000) != (co2 > 800 and light > 7000):
        triggered.add(7)
    if (co2 > 1100 and light > 7000) != (co2 > 1100 and light > 5000):
        triggered.add(8)
    if (co2 > 1100 and light > 7000) != (co2 > 1100 and light > 4000):
        triggered.add(9)
    if (co2 > 1100 and light > 7000) != (co2 > 1100 and light > 3000):
        triggered.add(10)
    if (co2 > 1100 and light > 7000) != (co2 > 1100 and light > 2000):
        triggered.add(11)

    if (co2 > 1250 and temp > 28) != (co2 > 950 and temp > 28):
        triggered.add(12)
    if (co2 > 1250 and temp > 28) != (co2 > 1250 and temp > 18):
        triggered.add(13)
    if (co2 > 1250 and temp > 28) != (co2 > 1000 and temp > 28):
        triggered.add(14)
    if (co2 > 1250 and temp > 28) != (co2 > 900 and temp > 28):
        triggered.add(15)
    if (co2 > 1250 and temp > 28) != (co2 > 1250 or temp > 28):
        triggered.add(16)
    if (co2 > 1250 and temp > 28) != (co2 > 1250 and temp <= 28):
        triggered.add(17)
    if (co2 > 1250 and temp > 28) != (co2 < 1250 and temp > 28):
        triggered.add(18)
    if (co2 > 1250 and temp > 28) != (light > 1250 and temp > 28):
        triggered.add(19)

    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1350 and light > 6000 and temp > 22):
        triggered.add(20)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 < 1150 and light > 6000 and temp > 22):
        triggered.add(21)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 7000 and temp > 22):
        triggered.add(22)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 8000 and temp > 22):
        triggered.add(23)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 4000 and temp > 22):
        triggered.add(24)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 6000 and temp > 32):
        triggered.add(25)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 6000 and temp > 12):
        triggered.add(26)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light < 6000 and temp > 22):
        triggered.add(27)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 6000 and temp < 22):
        triggered.add(28)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 or light > 6000 and temp > 22):
        triggered.add(29)
    if (co2 > 1150 and light > 6000 and temp > 22) != (co2 > 1150 and light > 6000 or temp > 22):
        triggered.add(30)

    if (co2 < 1050 and light > 6000) != (co2 < 1250 and light > 6000):
        triggered.add(31)
    if (co2 < 1050 and light > 6000) != (co2 < 1350 and light > 6000):
        triggered.add(32)
    if (co2 < 1050 and light > 6000) != (co2 < 1450 and light > 6000):
        triggered.add(33)
    if (co2 < 1050 and light > 6000) != (co2 > 1050 and light > 6000):
        triggered.add(34)
    if (co2 < 1050 and light > 6000) != (co2 < 1050 or light > 6000):
        triggered.add(35)
    if (co2 < 1050 and light > 6000) != (co2 < 1050 and light < 6000):
        triggered.add(36)
    if (co2 < 1050 and light > 6000) != (co2 < 1050 and light > 4000):
        triggered.add(37)
    if (co2 < 1050 and light > 6000) != (co2 < 1050 and light > 8000):
        triggered.add(38)

    if (temp > 28 and light > 6000) != (temp > 18 and light > 6000):
        triggered.add(39)
    if (temp > 28 and light > 6000) != (temp > 24 and light > 6000):
        triggered.add(40)
    if (temp > 28 and light > 6000) != (temp > 20 and light > 6000):
        triggered.add(41)
    if (temp > 28 and light > 6000) != (temp > 28 or light > 6000):
        triggered.add(42)
    if (temp > 28 and light > 6000) != (temp > 28 and light > 4000):
        triggered.add(43)
    if (temp > 28 and light > 6000) != (temp > 28 and light > 2000):
        triggered.add(44)
    if (temp > 28 and light > 6000) != (temp > 28 and light > 8000):
        triggered.add(45)

    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1300 and light > 6000 and temp > 22):
        triggered.add(46)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 600 and light > 6000 and temp > 22):
        triggered.add(47)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 300 and light > 6000 and temp > 22):
        triggered.add(48)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 < 1100 and light > 6000 and temp > 22):
        triggered.add(49)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1100 or light > 6000 and temp > 22):
        triggered.add(50)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1100 and light > 4000 and temp > 22):
        triggered.add(51)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1100 and light > 8000 and temp > 22):
        triggered.add(52)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1100 and light > 6000 or temp > 22):
        triggered.add(53)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1100 and light > 6000 and temp > 12):
        triggered.add(54)
    if (co2 > 1100 and light > 6000 and temp > 22) != (co2 > 1100 and light > 6000 and temp > 32):
        triggered.add(55)

    if (co2 < 1050 and light > 5000) != (co2 < 1250 and light > 5000):
        triggered.add(56)
    if (co2 < 1050 and light > 5000) != (co2 < 950 and light > 5000):
        triggered.add(57)
    if (co2 < 1050 and light > 5000) != (co2 < 1050 or light > 5000):
        triggered.add(58)
    if (co2 < 1050 and light > 5000) != (co2 < 1050 and light > 6000):
        triggered.add(59)
    if (co2 < 1050 and light > 5000) != (co2 < 1050 and light > 8000):
        triggered.add(60)
    if (co2 < 1050 and light > 5000) != (co2 < 1050 and light > 3000):
        triggered.add(61)

    if (3000 <= light <= 8000) != (2000 <= light <= 8000):
        triggered.add(62)
    if (3000 <= light <= 8000) != (1000 <= light <= 8000):
        triggered.add(63)
    if (3000 <= light <= 8000) != (4000 <= light <= 8000):
        triggered.add(64)
    if (3000 <= light <= 8000) != (6000 <= light <= 8000):
        triggered.add(65)

    if (co2 > 1180 and light > 5500) != (co2 > 1280 and light > 5500):
        triggered.add(66)
    if (co2 > 1180 and light > 5500) != (co2 > 1380 and light > 5500):
        triggered.add(67)
    if (co2 > 1180 and light > 5500) != (co2 < 1180 and light > 5500):
        triggered.add(68)
    if (co2 > 1180 and light > 5500) != (co2 > 1180 or light > 5500):
        triggered.add(69)
    if (co2 > 1180 and light > 5500) != (co2 > 1180 and light > 6600):
        triggered.add(70)
    if (co2 > 1180 and light > 5500) != (co2 > 1180 and light > 7700):
        triggered.add(71)
    if (co2 > 1180 and light > 5500) != (co2 > 1180 and light > 8800):
        triggered.add(72)

    if (light > 3500 and light < 8500) != (light > 5500 and light < 8500):
        triggered.add(73)
    if (light > 3500 and light < 8500) != (light > 6500 and light < 8500):
        triggered.add(74)
    if (light > 3500 and light < 8500) != (light > 7500 and light < 8500):
        triggered.add(75)
    if (light > 3500 and light < 8500) != (light < 3500 and light < 8500):
        triggered.add(76)
    if (light > 3500 and light < 8500) != (light > 3500 and light > 8500):
        triggered.add(77)
    if (light > 3500 and light < 8500) != (light > 3500 and light < 7500):
        triggered.add(78)
    if (light > 3500 and light < 8500) != (light > 3500 and light < 5500):
        triggered.add(79)
    if (light > 3500 and light < 8500) != (light > 3500 and light < 4500):
        triggered.add(80)

    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 900) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000):
        triggered.add(81)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1250) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000):
        triggered.add(82)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 3 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000):
        triggered.add(83)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 4500) ** 2 / 100 + (temp - 25) ** 2 < 10000):
        triggered.add(84)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 3500) ** 2 / 100 + (temp - 25) ** 2 < 10000):
        triggered.add(85)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 5500) ** 3 / 100 + (temp - 25) ** 2 < 10000):
        triggered.add(86)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 5500) ** 2 / 200 + (temp - 25) ** 2 < 10000):
        triggered.add(87)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 15) ** 2 < 10000):
        triggered.add(88)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 35) ** 2 < 10000):
        triggered.add(89)
    if ((co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 2 < 10000) != (
            (co2 - 1150) ** 2 + (light - 5500) ** 2 / 100 + (temp - 25) ** 3 < 10000):
        triggered.add(90)

    if (light / temp > 50 and light / temp < 300) != (light / temp > 60 and light / temp < 300):
        triggered.add(91)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 70 and light / temp < 300):
        triggered.add(92)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 80 and light / temp < 300):
        triggered.add(93)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 90 and light / temp < 300):
        triggered.add(94)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 50 or light / temp < 300):
        triggered.add(95)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 50 and light / temp < 200):
        triggered.add(96)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 50 and light / temp < 100):
        triggered.add(97)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 50 and light / temp < 350):
        triggered.add(98)
    if (light / temp > 50 and light / temp < 300) != (light / temp > 50 and light / temp < 400):
        triggered.add(99)

    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 1050) + abs(light - 5500) / 10 + abs(temp - 25) < 500):
        triggered.add(100)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 950) + abs(light - 5500) / 10 + abs(temp - 25) < 500):
        triggered.add(101)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 850) + abs(light - 5500) / 10 + abs(temp - 25) < 500):
        triggered.add(102)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 1150) + abs(light - 6500) / 10 + abs(temp - 25) < 500):
        triggered.add(103)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 1150) + abs(light - 7500) / 10 + abs(temp - 25) < 500):
        triggered.add(104)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 1150) + abs(light - 8500) / 20 + abs(temp - 25) < 500):
        triggered.add(105)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 15) < 500):
        triggered.add(106)
    if (abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 500) != (
            abs(co2 - 1150) + abs(light - 5500) / 10 + abs(temp - 25) < 400):
        triggered.add(107)

    if (light > 1200 and light < 9800) != (light > 2200 and light < 9800):
        triggered.add(108)
    if (light > 1200 and light < 9800) != (light > 4200 and light < 9800):
        triggered.add(109)
    if (light > 1200 and light < 9800) != (light > 6200 and light < 9800):
        triggered.add(110)
    if (light > 1200 and light < 9800) != (light > 7200 and light < 9800):
        triggered.add(111)
    if (light > 1200 and light < 9800) != (light > 1200 and light < 7800):
        triggered.add(112)
    if (light > 1200 and light < 9800) != (light > 1200 and light < 5800):
        triggered.add(113)
    if (light > 1200 and light < 9800) != (light > 1200 and light < 2800):
        triggered.add(114)
    if (light > 1200 and light < 9800) != (light > 1200 and light < 3800):
        triggered.add(115)
    if (light > 1200 and light < 9800) != (light > 1200 and light < 4800):
        triggered.add(116)

    if (temp > 3 and temp < 38) != (temp > 13 and temp < 38):
        triggered.add(117)
    if (temp > 3 and temp < 38) != (temp > 23 and temp < 38):
        triggered.add(118)
    if (temp > 3 and temp < 38) != (temp > 33 and temp < 38):
        triggered.add(119)
    if (temp > 3 and temp < 38) != (temp > 3 and temp < 8):
        triggered.add(120)
    if (temp > 3 and temp < 38) != (temp > 3 and temp < 18):
        triggered.add(121)
    if (temp > 3 and temp < 38) != (temp > 3 and temp < 28):
        triggered.add(122)
    if (temp > 3 and temp < 38) != (temp > 3 and temp < 30):
        triggered.add(123)

    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 1000 and light > 1500 and temp > 5):
        triggered.add(124)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 950 and light > 1500 and temp > 5):
        triggered.add(125)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 1250 and light > 1500 and temp > 5):
        triggered.add(126)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 or light > 1500 and temp > 5):
        triggered.add(127)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 3500 and temp > 5):
        triggered.add(128)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 5500 and temp > 5):
        triggered.add(129)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 6500 and temp > 5):
        triggered.add(130)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 1500 and temp > 15):
        triggered.add(131)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 1500 and temp > 25):
        triggered.add(132)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 1500 and temp > 35):
        triggered.add(133)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 1500 or temp > 5):
        triggered.add(134)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 2500 and temp > 5):
        triggered.add(135)
    if (co2 > 850 and light > 1500 and temp > 5) != (co2 > 850 and light > 4500 and temp > 5):
        triggered.add(136)

    return triggered


def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """"""
    generated_path = execute_validation_rules(particle[0], particle[1], particle[2])

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # : x:1-50, y:1-50, z:1-50
        self.bounds = bounds if bounds else [(1, 100), (1, 100), (1, 100)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """target pathsPSO"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(particles[i][0], particles[i][1], particles[i][2]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(gbest_particle[0], gbest_particle[1], gbest_particle[2]),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """Path PSO"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - Path ")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations")
    print(f"Path : {len(target_paths)}")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"Path {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "" if result['success'] else f"({result['best_fitness']:.3f})"
        print(f"{status} | {result['time']:.2f}s | iterations{result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f": {success_count}/{len(target_paths)} ({success_rate:.1f}%) | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs}")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations, {len(target_paths)}Path ")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"---  {run_idx}/{num_runs} ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f": {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs} runcompleted | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 1: 
    ws1 = wb.active
    ws1.title = ""
    ws1.sheet_view.showGridLines = False

    headers = ["Run", "", "", "", "Average Iterations", "(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f" {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # 2: Path 
    ws2 = wb.create_sheet(title="Path ")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Path ID", "", "", "", "Average Iterations", "Minimum Iterations", "Maximum Iterations"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"Path  {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # 3: 
    ws3 = wb.create_sheet(title="")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Path ", "", "(x,y,z)", "", "Iterations", "Path "]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"{run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # 4: target paths
    ws4 = wb.create_sheet(title="target paths")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Path ID", "target paths", ""]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"Path  {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f" : {filename}")
    print(f"{'=' * 70}")
    print(f":")
    print(f"  1.        - {len(all_results)} run")
    print(f"  2. Path        - Path ")
    print(f"  3.    -  runPath ")
    print(f"  4. target paths       - target paths")
    print(f"{'=' * 70}\n")

    return filename


def main():


    target_paths = [
        [1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20, 21, 22, 23, 25, 27, 28, 31, 32, 33, 34, 35, 45, 46, 49,
         52, 55, 56, 58, 66, 67, 68, 70, 71, 72, 74, 75, 76, 77, 79, 80, 81, 82, 83, 84, 85, 86, 96, 97, 111, 113, 114,
         115, 116, 119, 120, 121, 122, 123, 126, 130, 133],
        [1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20, 21, 22, 23, 25, 27, 28, 31, 32, 33, 34, 35, 45, 46, 49,
         52,
         55, 56, 58, 66, 67, 68, 70, 71, 72, 74, 75, 76, 77, 79, 80, 81, 83, 84, 85, 86, 97, 110, 111, 113, 114, 115,
         116,
         119, 120, 121, 122, 123, 126, 130, 133],
        [1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20, 21, 22, 23, 25, 27, 28, 31, 32, 33, 34, 35, 45, 46, 49,
         52,
         55, 56, 58, 68, 69, 74, 75, 76, 77, 79, 80, 81, 82, 84, 85, 86, 96, 97, 110, 111, 113, 114, 115, 116, 119, 120,
         121, 122, 123, 126, 130, 133],
        [8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20, 21, 22, 23, 27, 28, 31, 32, 33, 34, 35, 45, 46, 49, 52, 56, 58, 66,
         67,
         68, 70, 71, 72, 74, 75, 76, 77, 79, 80, 81, 83, 84, 85, 86, 90, 97, 110, 111, 113, 114, 115, 116, 120, 121,
         122,
         123, 126, 130],
        [1, 2, 3, 4, 8, 9, 10, 11, 26, 28, 29, 30, 31, 32, 33, 34, 35, 39, 41, 42, 50, 53, 54, 56, 58, 66, 67, 68, 70,
         71,
         72, 74, 75, 76, 77, 79, 80, 81, 83, 84, 85, 86, 96, 97, 110, 111, 113, 114, 115, 116, 118, 119, 120, 121, 126,
         130, 132, 133],
        [1, 2, 3, 4, 8, 9, 10, 11, 20, 21, 22, 23, 25, 27, 28, 31, 32, 33, 34, 35, 39, 40, 41, 42, 46, 49, 52, 55, 56,
         58,
         68, 69, 74, 75, 76, 77, 79, 80, 81, 82, 84, 85, 86, 96, 97, 110, 111, 113, 114, 115, 116, 119, 120, 121, 126,
         130,
         133],
        [1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 24, 27, 29, 30, 42, 43, 44, 50, 51, 53, 56, 58, 65, 66, 67,
         68,
         70, 71, 72, 74, 75, 76, 77, 79, 80, 81, 83, 84, 85, 86, 97, 110, 111, 113, 114, 115, 116, 120, 121, 122, 123,
         126,
         130],
        [1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20, 21, 22, 23, 25, 27, 28, 31, 32, 33, 34, 35, 45, 46, 49,
         52,
         55, 56, 58, 66, 67, 68, 71, 72, 75, 76, 77, 79, 80, 87, 96, 97, 111, 113, 114, 115, 116, 119, 120, 121, 122,
         126,
         133],
        [1, 2, 3, 4, 8, 9, 10, 11, 26, 28, 29, 30, 31, 32, 33, 34, 35, 39, 42, 50, 53, 54, 56, 58, 68, 69, 74, 75, 76,
         77,
         79, 80, 81, 82, 83, 84, 85, 86, 95, 98, 99, 110, 111, 113, 114, 115, 116, 118, 119, 120, 121, 126, 130, 132,
         133],
        [1, 2, 3, 4, 8, 9, 10, 11, 26, 28, 29, 30, 31, 32, 33, 34, 35, 42, 50, 53, 54, 56, 58, 66, 67, 68, 70, 71, 72,
         74,
         75, 76, 77, 79, 80, 81, 82, 83, 84, 85, 86, 95, 111, 113, 114, 115, 116, 118, 119, 120, 126, 130, 131, 132,
         133],
        [1, 2, 3, 4, 12, 14, 15, 16, 18, 19, 21, 29, 30, 31, 32, 33, 34, 35, 45, 47, 48, 49, 50, 53, 56, 58, 68, 69, 74,
         75, 76, 77, 79, 80, 81, 82, 84, 85, 86, 97, 110, 111, 113, 114, 115, 116, 119, 120, 121, 122, 123, 126, 130,
         133],
        [8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 20, 21, 22, 23, 25, 27, 28, 31, 32, 33, 34, 35, 45, 46, 49, 52, 55, 56,
         58,
         66, 67, 68, 70, 71, 72, 75, 76, 77, 79, 80, 96, 97, 102, 111, 113, 114, 115, 116, 119, 120, 121, 122, 126,
         133],
        [1, 2, 3, 4, 8, 9, 10, 11, 28, 29, 30, 31, 32, 33, 34, 35, 42, 50, 53, 56, 58, 66, 67, 68, 70, 71, 72, 74, 75,
         76,
         77, 79, 80, 81, 83, 84, 85, 86, 95, 110, 111, 113, 114, 115, 116, 117, 118, 119, 120, 126, 130, 131, 132, 133],
        [1, 2, 3, 4, 5, 6, 7, 16, 18, 19, 21, 29, 30, 34, 36, 38, 45, 47, 48, 49, 50, 53, 60, 68, 69, 76, 77, 78, 79,
         80,
         83, 96, 97, 100, 101, 102, 103, 104, 105, 112, 113, 114, 115, 116, 119, 120, 121, 122, 123, 124, 125, 126,
         133],
        [1, 2, 3, 4, 12, 14, 15, 16, 18, 19, 21, 29, 30, 31, 32, 33, 34, 35, 45, 47, 48, 49, 50, 53, 56, 58, 68, 69, 74,
         75, 76, 77, 79, 80, 81, 82, 84, 85, 86, 88, 90, 97, 111, 113, 114, 115, 116, 120, 121, 122, 123, 126, 130,
         133],
        [8, 9, 10, 11, 13, 16, 17, 20, 21, 22, 23, 25, 27, 28, 32, 33, 34, 35, 39, 41, 42, 46, 49, 52, 55, 58, 67, 68,
         70,
         71, 72, 74, 75, 76, 77, 79, 80, 82, 96, 97, 102, 110, 111, 113, 114, 115, 116, 119, 120, 121, 130, 132, 133],
        [8, 9, 10, 11, 26, 28, 29, 30, 31, 32, 33, 34, 35, 42, 50, 53, 54, 56, 58, 66, 67, 68, 70, 71, 72, 74, 75, 76,
         77,
         79, 80, 81, 83, 84, 85, 86, 89, 95, 99, 111, 113, 114, 115, 116, 118, 119, 120, 121, 126, 130, 132, 133],
        [8, 9, 10, 11, 13, 16, 17, 21, 22, 23, 25, 27, 28, 34, 35, 39, 41, 42, 49, 52, 55, 58, 68, 70, 71, 72, 74, 75,
         76,
         77, 79, 80, 96, 97, 100, 101, 102, 107, 111, 113, 114, 115, 116, 118, 119, 120, 121, 130, 132, 133],
        [1, 2, 3, 4, 5, 6, 7, 16, 18, 19, 21, 29, 30, 34, 36, 38, 45, 47, 48, 49, 50, 53, 60, 68, 69, 75, 76, 77, 79,
         80,
         83, 96, 97, 100, 101, 102, 103, 104, 105, 113, 114, 115, 116, 119, 120, 121, 122, 123, 127, 134],
        [1, 2, 3, 4, 5, 6, 7, 15, 16, 18, 19, 21, 29, 30, 34, 36, 38, 45, 47, 48, 49, 50, 53, 60, 68, 69, 76, 77, 78,
         79,
         80, 83, 96, 97, 106, 107, 112, 113, 114, 115, 116, 119, 120, 121, 122, 123, 124, 125, 126, 133],
        [1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15, 16, 18, 19, 24, 27, 29, 30, 42, 43, 44, 50, 51, 53, 56, 58, 65, 73, 74,
         75,
         76, 77, 80, 81, 85, 97, 110, 111, 114, 115, 116, 119, 120, 121, 122, 123, 126, 129, 130, 133],
        [1, 2, 3, 4, 12, 15, 16, 18, 19, 21, 29, 30, 34, 36, 38, 45, 47, 48, 49, 50, 53, 57, 60, 68, 69, 74, 75, 76, 77,
         79, 80, 81, 83, 96, 97, 110, 111, 113, 114, 115, 116, 119, 120, 121, 122, 124, 126, 130, 133],
        [1, 2, 3, 4, 9, 10, 11, 12, 14, 15, 16, 18, 19, 24, 27, 29, 30, 42, 43, 44, 50, 51, 53, 65, 69, 73, 74, 75, 76,
         77, 84, 85, 86, 87, 97, 110, 111, 114, 115, 119, 120, 121, 122, 123, 126, 129, 130, 133, 136],
        [1, 2, 3, 4, 9, 10, 11, 12, 14, 15, 16, 18, 19, 24, 27, 29, 30, 42, 43, 44, 50, 51, 53, 65, 73, 74, 75, 76, 77,
         84, 85, 86, 97, 109, 110, 111, 114, 115, 119, 120, 121, 122, 123, 126, 129, 130, 133, 136],
        [1, 2, 3, 4, 12, 14, 15, 16, 18, 19, 27, 29, 30, 42, 50, 53, 63, 69, 76, 86, 91, 92, 93, 94, 100, 101, 102, 103,
         104, 107, 108, 109, 110, 111, 119, 120, 121, 122, 123, 126, 128, 129, 130, 133, 135, 136],
        [1, 2, 3, 4, 16, 18, 19, 30, 35, 36, 37, 42, 43, 44, 53, 58, 61, 65, 73, 74, 75, 76, 77, 83, 86, 97, 103, 104,
         105, 107, 110, 111, 114, 115, 119, 120, 121, 122, 123, 124, 125, 126, 129, 130, 133, 136],
        [1, 2, 3, 4, 12, 15, 16, 18, 19, 30, 35, 36, 42, 44, 53, 58, 62, 63, 76, 83, 86, 92, 93, 94, 100, 101, 102, 105,
         108, 109, 110, 111, 119, 120, 121, 122, 123, 124, 126, 128, 129, 130, 133, 135, 136],
        [1, 2, 3, 4, 16, 18, 19, 30, 35, 36, 42, 44, 53, 58, 61, 64, 65, 73, 74, 75, 76, 77, 83, 86, 97, 103, 104, 105,
         107, 109, 110, 111, 114, 115, 120, 121, 122, 123, 124, 125, 126, 129, 130, 133, 136],
        [1, 2, 3, 4, 15, 16, 18, 19, 30, 35, 36, 37, 42, 43, 44, 53, 59, 60, 65, 68, 69, 74, 75, 76, 77, 79, 80, 81, 83,
         96, 97, 110, 111, 113, 114, 115, 116, 119, 120, 121, 122, 124, 125, 126, 130, 133]
    ]

    print("=" * 70)
    print("baseline PSO")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("Program completed")


if __name__ == "__main__":
    main()