import os
import random
from collections import deque
import numpy as np
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# --- 全局状态范围配置 ---
STATE_MIN_X, STATE_MAX_X = 1000, 10000
STATE_MIN_Y, STATE_MAX_Y = 38, 85
STATE_MIN_Z, STATE_MAX_Z = 38, 85

def generate_input():
    return [
        random.randint(STATE_MIN_X, STATE_MAX_X),
        random.randint(STATE_MIN_Y, STATE_MAX_Y),
        random.randint(STATE_MIN_Z, STATE_MAX_Z)
    ]


def execute_Tr(light, moisture, humidity):
    """分析正常条件下的分支覆盖"""
    actions = []
    triggered = set()

    if (light > 2000 and light < 8500) != (light > 3000 and light < 8500):
        triggered.add(1)
    if (light > 2000 and light < 8500) != (light > 4000 and light < 8500):
        triggered.add(2)
    if (light > 2000 and light < 8500) != (light < 2000 and light < 8500):
        triggered.add(3)
    if (light > 2000 and light < 8500) != (light > 2000 or light < 8500):
        triggered.add(4)
    if (light > 2000 and light < 8500) != (light > 2000 and light < 9500):
        triggered.add(5)

    if (light < 7500 and light > 1500) != (light < 7500 and light > 2500):
        triggered.add(6)
    if (light < 7500 and light > 1500) != (light < 7500 and light > 3500):
        triggered.add(7)
    if (light < 7500 and light > 1500) != (light < 7500 and light > 4500):
        triggered.add(8)

    if (moisture > 38 and moisture < 62) != (moisture > 58 and moisture < 62):
        triggered.add(9)
    if (moisture > 38 and moisture < 62) != (moisture > 48 and moisture < 62):
        triggered.add(10)
    if (moisture > 38 and moisture < 62) != (moisture > 38 or moisture < 62):
        triggered.add(11)
    if (moisture > 38 and moisture < 62) != (moisture > 38 and moisture < 42):
        triggered.add(12)
    if (moisture > 38 and moisture < 62) != (moisture > 38 and moisture < 52):
        triggered.add(13)

    if (humidity > 30 and humidity < 60) != (humidity > 50 and humidity < 60):
        triggered.add(14)
    if (humidity > 30 and humidity < 60) != (humidity > 40 and humidity < 60):
        triggered.add(15)
    if (humidity > 30 and humidity < 60) != (humidity > 30 and humidity < 70):
        triggered.add(16)
    if (humidity > 30 and humidity < 60) != (humidity > 30 and humidity < 75):
        triggered.add(17)
    if (humidity > 30 and humidity < 60) != (humidity > 30 or humidity < 60):
        triggered.add(18)
    if (humidity > 30 and humidity < 60) != (humidity > 30 and humidity < 40):
        triggered.add(19)

    if (2000 <= light <= 6000 and 50 <= moisture <= 70 and 40 <= humidity <= 75) != (
            4000 <= light <= 6000 and 50 <= moisture <= 70 and 40 <= humidity <= 75):
        triggered.add(20)
    if (2000 <= light <= 6000 and 50 <= moisture <= 70 and 40 <= humidity <= 75) != (
            2000 <= light <= 9000 and 50 <= moisture <= 70 and 40 <= humidity <= 75):
        triggered.add(21)
    if (2000 <= light <= 6000 and 50 <= moisture <= 70 and 40 <= humidity <= 75) != (
            2000 <= light <= 8000 and 50 <= moisture <= 70 and 40 <= humidity <= 75):
        triggered.add(22)
    if (2000 <= light <= 6000 and 50 <= moisture <= 70 and 40 <= humidity <= 75) != (
            2000 <= light <= 6000 and 40 <= moisture <= 70 and 40 <= humidity <= 75):
        triggered.add(23)
    if (2000 <= light <= 6000 and 50 <= moisture <= 70 and 40 <= humidity <= 75) != (
            2000 <= light <= 6000 and 50 <= moisture <= 70 and 30 <= humidity <= 75):
        triggered.add(24)
    if (2000 <= light <= 6000 and 50 <= moisture <= 70 and 40 <= humidity <= 75) != (
            2000 <= light <= 6000 and 40 <= moisture <= 70 and 40 <= humidity <= 75):
        triggered.add(25)
    if (2000 <= light <= 6000 and 50 <= moisture <= 70 and 40 <= humidity <= 75) != (
            2000 <= light <= 6000 and 50 <= moisture <= 70 or 40 <= humidity <= 75):
        triggered.add(26)

    if (light > 5500 and moisture > 68 and humidity > 65) != (
            light > 7500 and moisture > 68 and humidity > 65):
        triggered.add(27)
    if (light > 5500 and moisture > 68 and humidity > 65) != (
            light > 8500 and moisture > 68 and humidity > 65):
        triggered.add(28)
    if (light > 5500 and moisture > 68 and humidity > 65) != (
            light < 5500 and moisture > 68 and humidity > 65):
        triggered.add(29)
    if (light > 5500 and moisture > 68 and humidity > 65) != (
            light > 5500 and moisture > 48 and humidity > 65):
        triggered.add(30)
    if (light > 5500 and moisture > 68 and humidity > 65) != (
            light > 5500 and moisture > 38 and humidity > 65):
        triggered.add(31)
    if (light > 5500 and moisture > 68 and humidity > 65) != (
            light > 5500 and moisture > 68 and humidity < 65):
        triggered.add(32)
    if (light > 5500 and moisture > 68 and humidity > 65) != (
            light > 5500 and moisture > 68 and humidity > 35):
        triggered.add(33)
    if (light > 5500 and moisture > 68 and humidity > 65) != (
            light > 5500 and moisture > 68 or humidity > 65):
        triggered.add(34)

    if (light > 7500 and moisture > 28 and humidity > 45) != (
            light > 4500 and moisture > 28 and humidity > 45):
        triggered.add(35)
    if (light > 7500 and moisture > 28 and humidity > 45) != (
            light > 2500 and moisture > 28 and humidity > 45):
        triggered.add(36)
    if (light > 7500 and moisture > 28 and humidity > 45) != (
            light < 7500 and moisture > 28 and humidity > 45):
        triggered.add(37)
    if (light > 7500 and moisture > 28 and humidity > 45) != (
            light > 7500 and moisture > 68 and humidity > 45):
        triggered.add(38)
    if (light > 7500 and moisture > 28 and humidity > 45) != (
            light > 7500 and moisture > 48 and humidity > 45):
        triggered.add(39)
    if (light > 7500 and moisture > 28 and humidity > 45) != (
            light > 7500 and moisture > 28 and humidity < 45):
        triggered.add(40)
    if (light > 7500 and moisture > 28 and humidity > 45) != (
            light > 7500 and moisture > 28 and humidity > 75):
        triggered.add(41)
    if (light > 7500 and moisture > 28 and humidity > 45) != (
            light > 7500 and moisture > 28 and humidity > 65):
        triggered.add(42)

    if (4500 <= light <= 6500 and abs(moisture - humidity) > 25) != (
            2500 <= light <= 6500 and abs(moisture - humidity) > 25):
        triggered.add(43)
    if (4500 <= light <= 6500 and abs(moisture - humidity) > 25) != (
            2500 <= light <= 6500 and abs(moisture - humidity) > 25):
        triggered.add(44)
    if (4500 <= light <= 6500 and abs(moisture - humidity) > 25) != (
            4500 <= light <= 9500 and abs(moisture - humidity) > 25):
        triggered.add(45)
    if (4500 <= light <= 6500 and abs(moisture - humidity) > 25) != (
            4500 <= light <= 8500 and abs(moisture - humidity) > 25):
        triggered.add(46)
    if (4500 <= light <= 6500 and abs(moisture - humidity) > 25) != (
            4500 <= light <= 6500 and abs(moisture + humidity) > 25):
        triggered.add(47)
    if (4500 <= light <= 6500 and abs(moisture - humidity) > 25) != (
            4500 <= light <= 6500 and abs(moisture - humidity + 250) > 25):
        triggered.add(48)
    if (4500 <= light <= 6500 and abs(moisture - humidity) > 25) != (
            4500 <= light <= 6500 and abs(moisture - humidity) > 5):
        triggered.add(49)
    if (4500 <= light <= 6500 and abs(moisture - humidity) > 25) != (
            4500 <= light <= 6500 and abs(moisture - humidity) > 10):
        triggered.add(50)

    if (2500 <= light <= 5500 and 38 <= moisture <= 68 and humidity < 45) != (
            1500 <= light <= 5500 and 38 <= moisture <= 68 and humidity < 45):
        triggered.add(51)
    if (2500 <= light <= 5500 and 38 <= moisture <= 68 and humidity < 45) != (
            2500 <= light <= 7500 and 38 <= moisture <= 68 and humidity < 45):
        triggered.add(52)
    if (2500 <= light <= 5500 and 38 <= moisture <= 68 and humidity < 45) != (
            2500 <= light <= 8500 and 38 <= moisture <= 68 and humidity < 45):
        triggered.add(53)
    if (2500 <= light <= 5500 and 38 <= moisture <= 68 and humidity < 45) != (
            2500 <= light <= 5500 and 48 <= moisture <= 68 and humidity < 45):
        triggered.add(54)
    if (2500 <= light <= 5500 and 38 <= moisture <= 68 and humidity < 45) != (
            2500 <= light <= 5500 and 38 <= moisture <= 85 and humidity < 45):
        triggered.add(55)
    if (2500 <= light <= 5500 and 38 <= moisture <= 68 and humidity < 45) != (
            2500 <= light <= 5500 and 38 <= moisture <= 68 and humidity < 55):
        triggered.add(56)
    if (2500 <= light <= 5500 and 38 <= moisture <= 68 and humidity < 45) != (
            2500 <= light <= 5500 and 38 <= moisture <= 68 and humidity < 75):
        triggered.add(57)
    if (2500 <= light <= 5500 and 38 <= moisture <= 68 and humidity < 45) != (
            2500 <= light <= 5500 and 38 <= moisture <= 68 or humidity < 45):
        triggered.add(58)
    if (2500 <= light <= 5500 and 38 <= moisture <= 68 and humidity < 45) != (
            2500 <= light <= 5500 or 38 <= moisture <= 68 and humidity < 45):
        triggered.add(59)

    if (light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70) != (
            light > 1000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70):
        triggered.add(60)
    if (light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70) != (
            light > 3000 and light < 10000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70):
        triggered.add(61)
    if (light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70) != (
            light > 3000 and light < 7000 or moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70):
        triggered.add(62)
    if (light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70) != (
            light > 3000 and light < 7000 and moisture > 40 and moisture < 75 and humidity > 55 and humidity < 70):
        triggered.add(63)
    if (light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70) != (
            light > 3000 and light < 7000 and moisture > 50 and moisture < 75 and humidity > 55 and humidity < 70):
        triggered.add(64)
    if (light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70) != (
            light > 3000 and light < 7000 and moisture > 60 and moisture < 85 and humidity > 55 and humidity < 70):
        triggered.add(65)
    if (light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70) != (
            light > 3000 and light < 7000 and moisture > 60 or moisture < 75 and humidity > 55 and humidity < 70):
        triggered.add(66)
    if (light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70) != (
            light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 45 and humidity < 70):
        triggered.add(67)
    if (light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70) != (
            light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 35 and humidity < 70):
        triggered.add(68)
    if (light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70) != (
            light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 80):
        triggered.add(69)
    if (light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70) != (
            light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 or humidity < 70):
        triggered.add(70)
    if (light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70) != (
            light > 1000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70):
        triggered.add(71)

    if (1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 60) != (
            1800 <= light <= 9200 and 38 <= moisture <= 66 and 46 <= humidity <= 60):
        triggered.add(72)
    if (1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 60) != (
            1800 <= light <= 5900 and 38 <= moisture <= 66 and 46 <= humidity <= 60):
        triggered.add(73)
    if (1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 60) != (
            1800 <= light <= 6200 and 48 <= moisture <= 66 and 46 <= humidity <= 60):
        triggered.add(74)
    if (1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 60) != (
            1800 <= light <= 6200 and 58 <= moisture <= 66 and 46 <= humidity <= 60):
        triggered.add(75)
    if (1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 60) != (
            1800 <= light <= 6200 and 38 <= moisture <= 76 and 46 <= humidity <= 60):
        triggered.add(76)
    if (1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 60) != (
            1800 <= light <= 6200 or 38 <= moisture <= 66 and 46 <= humidity <= 60):
        triggered.add(77)
    if (1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 60) != (
            1800 <= light <= 6200 and 38 <= moisture <= 66 and 36 <= humidity <= 60):
        triggered.add(78)
    if (1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 60) != (
            1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 50):
        triggered.add(79)
    if (1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 60) != (
            1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 70):
        triggered.add(80)
    if (1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 60) != (
            1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 80):
        triggered.add(81)
    if (1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 60) != (
            1800 <= light <= 6200 and 38 <= moisture <= 66 or 46 <= humidity <= 60):
        triggered.add(82)
    if (1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 60) != (
            1800 <= light <= 8200 and 38 <= moisture <= 66 and 46 <= humidity <= 60):
        triggered.add(83)

    if (light > 6500 and moisture > 70 and humidity < 48) != (
            light > 5500 and moisture > 70 and humidity < 48):
        triggered.add(84)
    if (light > 6500 and moisture > 70 and humidity < 48) != (
            light > 4500 and moisture > 70 and humidity < 48):
        triggered.add(85)
    if (light > 6500 and moisture > 70 and humidity < 48) != (
            light > 8500 and moisture > 70 and humidity < 48):
        triggered.add(86)
    if (light > 6500 and moisture > 70 and humidity < 48) != (
            light > 6500 and moisture > 40 and humidity < 48):
        triggered.add(87)
    if (light > 6500 and moisture > 70 and humidity < 48) != (
            light > 6500 and moisture > 20 and humidity < 48):
        triggered.add(88)
    if (light > 6500 and moisture > 70 and humidity < 48) != (
            light > 6500 and moisture > 70 and humidity < 38):
        triggered.add(89)
    if (light > 6500 and moisture > 70 and humidity < 48) != (
            light > 6500 and moisture > 70 and humidity < 28):
        triggered.add(90)
    if (light > 6500 and moisture > 70 and humidity < 48) != (
            light > 6500 or moisture > 70 and humidity < 48):
        triggered.add(91)
    if (light > 6500 and moisture > 70 and humidity < 48) != (
            light > 6500 and moisture > 70 or humidity < 48):
        triggered.add(92)
    if (light > 6500 and moisture > 70 and humidity < 48) != (
            light > 6500 and moisture > 70 and humidity > 48):
        triggered.add(93)

    if (light < 6800 and moisture < 42) != (light < 4800 and moisture < 42):
        triggered.add(94)
    if (light < 6800 and moisture < 42) != (light < 9800 and moisture < 42):
        triggered.add(95)
    if (light < 6800 and moisture < 42) != (light > 3800 and moisture < 42):
        triggered.add(96)
    if (light < 6800 and moisture < 42) != (light < 6800 or moisture < 42):
        triggered.add(97)
    if (light < 6800 and moisture < 42) != (light < 6800 and moisture < 62):
        triggered.add(98)
    if (light < 6800 and moisture < 42) != (light < 6800 and moisture < 72):
        triggered.add(99)
    if (light < 6800 and moisture < 42) != (light < 6800 and moisture < 82):
        triggered.add(100)

    return triggered


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0


# Fixed: Proper Python list syntax (removed Chinese labels)
TARGET_PATHS = [
    [3, 9, 10, 12, 16, 17, 18, 23, 25, 26, 34, 35, 36, 37, 47, 48, 49, 50, 57, 58, 59, 62, 63, 66, 70, 77, 80,
     81, 82, 97, 98, 99, 100],
    [2, 3, 8, 12, 13, 16, 17, 18, 20, 34, 36, 37, 57, 58, 59, 62, 63, 64, 66, 70, 77, 80, 81, 82, 97, 98, 99, 100],
    [3, 9, 10, 12, 19, 23, 25, 26, 35, 36, 37, 47, 48, 49, 50, 56, 57, 58, 59, 62, 70, 74, 75, 79, 97, 98, 99, 100],
    [2, 3, 8, 9, 12, 16, 17, 18, 23, 25, 26, 36, 37, 57, 58, 59, 62, 63, 66, 70, 77, 80, 81, 82, 97, 98, 99, 100],
    [3, 9, 12, 13, 16, 17, 18, 21, 22, 26, 30, 31, 34, 35, 36, 37, 47, 48, 49, 50, 62, 63, 64, 66, 70, 97, 98, 99,
     100],
    [3, 9, 12, 16, 17, 18, 21, 22, 26, 35, 36, 37, 47, 48, 49, 50, 62, 63, 66, 70, 77, 80, 81, 82, 97, 98, 99, 100],
    [3, 9, 12, 16, 17, 18, 21, 22, 26, 30, 31, 34, 35, 36, 37, 47, 48, 49, 50, 62, 63, 64, 66, 70, 97, 98, 99, 100],
    [3, 9, 12, 13, 16, 17, 18, 21, 22, 26, 35, 36, 37, 62, 63, 64, 66, 70, 72, 77, 82, 83, 91, 97, 98, 99, 100],
    [3, 11, 14, 19, 26, 32, 33, 34, 35, 36, 37, 47, 48, 49, 50, 62, 66, 67, 68, 70, 76, 77, 82, 97, 100],
    [1, 2, 3, 7, 8, 9, 10, 12, 19, 23, 25, 26, 36, 37, 56, 57, 58, 59, 70, 74, 75, 79, 97, 98, 99, 100],
    [3, 9, 12, 13, 14, 15, 24, 26, 47, 48, 49, 50, 52, 53, 58, 59, 62, 70, 77, 78, 82, 92, 97, 98, 99, 100],
    [3, 12, 13, 18, 26, 34, 35, 36, 37, 47, 48, 49, 50, 58, 59, 62, 66, 69, 70, 77, 81, 82, 97, 98, 99, 100],
    [1, 2, 3, 6, 7, 8, 9, 12, 13, 14, 15, 24, 26, 51, 58, 59, 70, 77, 78, 82, 92, 97, 98, 99, 100],
    [3, 9, 12, 13, 19, 21, 22, 26, 35, 36, 37, 47, 48, 62, 63, 64, 66, 70, 73, 75, 79, 97, 98, 99, 100],
    [3, 9, 10, 19, 23, 25, 26, 35, 36, 37, 47, 48, 49, 50, 56, 57, 58, 59, 62, 70, 74, 75, 79, 94],
    [3, 11, 16, 17, 18, 26, 27, 28, 29, 32, 35, 36, 37, 62, 65, 66, 70, 91, 92, 93, 97, 100],
    [3, 11, 14, 19, 26, 32, 33, 34, 35, 36, 37, 62, 66, 70, 77, 82, 84, 85, 91, 92, 97, 100],
    [3, 4, 6, 7, 8, 12, 13, 16, 17, 18, 26, 37, 60, 62, 66, 70, 71, 79, 97, 98, 99, 100],
    [3, 9, 10, 12, 14, 19, 23, 25, 26, 47, 48, 54, 62, 70, 77, 78, 82, 92, 97, 98, 99, 100],
    [4, 5, 9, 10, 16, 17, 18, 26, 31, 34, 37, 38, 39, 40, 41, 45, 66, 70, 91, 95, 96, 97],
    [4, 5, 9, 12, 13, 14, 19, 21, 26, 37, 38, 40, 41, 42, 70, 72, 77, 82, 87, 88, 91, 92],
    [2, 3, 8, 11, 14, 19, 20, 43, 44, 55, 58, 59, 62, 66, 68, 70, 77, 92, 97, 99, 100],
    [3, 11, 14, 15, 32, 33, 34, 45, 46, 58, 62, 66, 68, 70, 86, 89, 90, 93, 97, 100],
    [4, 11, 16, 17, 18, 26, 32, 33, 34, 37, 40, 41, 42, 61, 62, 66, 70, 91, 92, 93]
]


class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 2000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(TARGET_PATHS)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }


def compute_robustness(state, path):
    base = execute_Tr(state[0], state[1], state[2])
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                    np.clip(state[0] + dx, STATE_MIN_X, STATE_MAX_X),
                    np.clip(state[1] + dy, STATE_MIN_Y, STATE_MAX_Y),
                    np.clip(state[2] + dz, STATE_MIN_Z, STATE_MAX_Z)
                ])
                n_trig = execute_Tr(neighbor[0], neighbor[1], neighbor[2])
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = TARGET_PATHS[target_path_idx]
    samples = []
    attempts = 0

    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(STATE_MIN_X, STATE_MAX_X),
            random.randint(STATE_MIN_Y, STATE_MAX_Y),
            random.randint(STATE_MIN_Z, STATE_MAX_Z)
        ])
        triggered = execute_Tr(state[0], state[1], state[2])

        if not triggered:
            continue

        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)

        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }

        samples.append(sample_data)

    return samples


def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample[
                'robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)

        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]

    return selected_samples


def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = TARGET_PATHS[target_path_idx]
        samples = []
        attempts = 0

        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(STATE_MIN_X, STATE_MAX_X),
                random.randint(STATE_MIN_Y, STATE_MAX_Y),
                random.randint(STATE_MIN_Z, STATE_MAX_Z)
            ])
            triggered = execute_Tr(state[0], state[1], state[2])

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)

            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)

        random.shuffle(samples)
        return samples[:config.top_k_samples]

    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")

        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)


def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}

    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates

    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}

        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config,
                                                         shared_candidates[path_idx])

            strategy_results[path_idx] = samples

        results[strategy_name] = strategy_results

    return results


def analyze_fitness_values(results, config):
    analysis_results = {}

    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []

        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])

        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }

        all_scores = []

        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])

        scores_array = np.array(all_scores)

        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)

        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })

        analysis_results[strategy_name] = analysis

    return analysis_results


def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []

    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)

    df = pd.DataFrame(df_data)
    return df, analysis_results


def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []

    print(f"Starting {num_runs} experiments...")

    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")

        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)

        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)

        print(f"Experiment {run_idx + 1} completed.")

    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df


def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')

    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width

    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []

    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]

        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)

    stats_df = pd.DataFrame(stats_data)

    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Results saved to: {output_path}")


def main():
    results_df = run_multiple_experiments(num_runs=20)

    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")

    save_results_to_excel(results_df, output_path)

    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)

    return results_df, output_path


if __name__ == "__main__":
    results, output_file = main()