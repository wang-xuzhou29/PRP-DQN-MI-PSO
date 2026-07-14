
import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
import numpy as np
import random
import time
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
from datetime import datetime

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# ===  ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'ACTION_DIM': 3,
    'MIN_VALUES': np.array([1, 1, 1]),
    'MAX_VALUES': np.array([128, 200, 255]),
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 3,
    'PPO_BATCH_SIZE': 64,
    'PPO_EPOCHS': 4,
    'CLIP_EPSILON': 0.2,
    'GAMMA': 0.99,
    'GAE_LAMBDA': 0.95,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'ACTOR_LR': 3e-4,
    'CRITIC_LR': 3e-4,
    'NUM_RUNS': 20,
    'TOP_K_SAMPLES': 20,
    'TARGET_PATHS': [
        [3, 9, 10, 12, 16, 17, 18, 23, 25, 26, 34, 35, 36, 37, 47, 48, 49, 50, 57, 58, 59, 62, 63, 66, 70, 77, 80,
         81, 82, 97, 98, 99, 100],
        [2, 3, 8, 12, 13, 16, 17, 18, 20, 34, 36, 37, 57, 58, 59, 62, 63, 64, 66, 70, 77, 80, 81, 82, 97, 98, 99, 100],
        [3, 9, 10, 12, 19, 23, 25, 26, 35, 36, 37, 47, 48, 49, 50, 56, 57, 58, 59, 62, 70, 74, 75, 79, 97, 98, 99, 100],
        [2, 3, 8, 9, 12, 16, 17, 18, 23, 25, 26, 36, 37, 57, 58, 59, 62, 63, 66, 70, 77, 80, 81, 82, 97, 98, 99, 100],
        [3, 9, 12, 13, 16, 17, 18, 21, 22, 26, 30, 31, 34, 35, 36, 37, 47, 48, 49, 50, 62, 63, 64, 66, 70, 97, 98, 99,
         100],
        [3, 9, 12, 16, 17, 18, 21, 22, 26, 35, 36, 37, 47, 48, 49, 50, 62, 63, 66, 70, 77, 80, 81, 82, 97, 98, 99, 100],
        [3, 9, 12, 16, 17, 18, 21, 22, 26, 30, 31, 34, 35, 36, 37, 47, 48, 49, 50, 62, 63, 64, 66, 70, 97, 98, 99, 100],
        [3, 9, 12, 13, 16, 17, 18, 21, 22, 26, 35, 36, 37, 62, 63, 64, 66, 70, 72, 77, 82, 83, 91, 97, 98, 99, 100],
        [3, 11, 14, 19, 26, 32, 33, 34, 35, 36, 37, 47, 48, 49, 50, 62, 66, 67, 68, 70, 76, 77, 82, 97, 100],
        [1, 2, 3, 7, 8, 9, 10, 12, 19, 23, 25, 26, 36, 37, 56, 57, 58, 59, 70, 74, 75, 79, 97, 98, 99, 100],
        [3, 9, 12, 13, 14, 15, 24, 26, 47, 48, 49, 50, 52, 53, 58, 59, 62, 70, 77, 78, 82, 92, 97, 98, 99, 100],
        [3, 12, 13, 18, 26, 34, 35, 36, 37, 47, 48, 49, 50, 58, 59, 62, 66, 69, 70, 77, 81, 82, 97, 98, 99, 100],
        [1, 2, 3, 6, 7, 8, 9, 12, 13, 14, 15, 24, 26, 51, 58, 59, 70, 77, 78, 82, 92, 97, 98, 99, 100],
        [3, 9, 12, 13, 19, 21, 22, 26, 35, 36, 37, 47, 48, 62, 63, 64, 66, 70, 73, 75, 79, 97, 98, 99, 100],
        [3, 9, 10, 19, 23, 25, 26, 35, 36, 37, 47, 48, 49, 50, 56, 57, 58, 59, 62, 70, 74, 75, 79, 94],
        [3, 11, 16, 17, 18, 26, 27, 28, 29, 32, 35, 36, 37, 62, 65, 66, 70, 91, 92, 93, 97, 100],
        [3, 11, 14, 19, 26, 32, 33, 34, 35, 36, 37, 62, 66, 70, 77, 82, 84, 85, 91, 92, 97, 100],
        [3, 4, 6, 7, 8, 12, 13, 16, 17, 18, 26, 37, 60, 62, 66, 70, 71, 79, 97, 98, 99, 100],
        [3, 9, 10, 12, 14, 19, 23, 25, 26, 47, 48, 54, 62, 70, 77, 78, 82, 92, 97, 98, 99, 100],
        [4, 5, 9, 10, 16, 17, 18, 26, 31, 34, 37, 38, 39, 40, 41, 45, 66, 70, 91, 95, 96, 97],
        [4, 5, 9, 12, 13, 14, 19, 21, 26, 37, 38, 40, 41, 42, 70, 72, 77, 82, 87, 88, 91, 92],
        [2, 3, 8, 11, 14, 19, 20, 43, 44, 55, 58, 59, 62, 66, 68, 70, 77, 92, 97, 99, 100],
        [3, 11, 14, 15, 32, 33, 34, 45, 46, 58, 62, 66, 68, 70, 86, 89, 90, 93, 97, 100],
        [4, 11, 16, 17, 18, 26, 32, 33, 34, 37, 40, 41, 42, 61, 62, 66, 70, 91, 92, 93]
    ],
}

# ===  ===
def execute_Tr(light, moisture, humidity):
    """分析正常条件下的分支覆盖"""
    actions = []
    triggered = set()

    if (light > 2000 and light < 8500) != (light > 3000 and light < 8500):
        triggered.add(1)
    if (light > 2000 and light < 8500) != (light > 4000 and light < 8500):
        triggered.add(2)
    if (light > 2000 and light < 8500) != (light < 2000 and light < 8500):
        triggered.add(3)
    if (light > 2000 and light < 8500) != (light > 2000 or light < 8500):
        triggered.add(4)
    if (light > 2000 and light < 8500) != (light > 2000 and light < 9500):
        triggered.add(5)

    if (light < 7500 and light > 1500) != (light < 7500 and light > 2500):
        triggered.add(6)
    if (light < 7500 and light > 1500) != (light < 7500 and light > 3500):
        triggered.add(7)
    if (light < 7500 and light > 1500) != (light < 7500 and light > 4500):
        triggered.add(8)

    if (moisture > 38 and moisture < 62) != (moisture > 58 and moisture < 62):
        triggered.add(9)
    if (moisture > 38 and moisture < 62) != (moisture > 48 and moisture < 62):
        triggered.add(10)
    if (moisture > 38 and moisture < 62) != (moisture > 38 or moisture < 62):
        triggered.add(11)
    if (moisture > 38 and moisture < 62) != (moisture > 38 and moisture < 42):
        triggered.add(12)
    if (moisture > 38 and moisture < 62) != (moisture > 38 and moisture < 52):
        triggered.add(13)

    if (humidity > 30 and humidity < 60) != (humidity > 50 and humidity < 60):
        triggered.add(14)
    if (humidity > 30 and humidity < 60) != (humidity > 40 and humidity < 60):
        triggered.add(15)
    if (humidity > 30 and humidity < 60) != (humidity > 30 and humidity < 70):
        triggered.add(16)
    if (humidity > 30 and humidity < 60) != (humidity > 30 and humidity < 75):
        triggered.add(17)
    if (humidity > 30 and humidity < 60) != (humidity > 30 or humidity < 60):
        triggered.add(18)
    if (humidity > 30 and humidity < 60) != (humidity > 30 and humidity < 40):
        triggered.add(19)

    if (2000 <= light <= 6000 and 50 <= moisture <= 70 and 40 <= humidity <= 75) != (
            4000 <= light <= 6000 and 50 <= moisture <= 70 and 40 <= humidity <= 75):
        triggered.add(20)
    if (2000 <= light <= 6000 and 50 <= moisture <= 70 and 40 <= humidity <= 75) != (
            2000 <= light <= 9000 and 50 <= moisture <= 70 and 40 <= humidity <= 75):
        triggered.add(21)
    if (2000 <= light <= 6000 and 50 <= moisture <= 70 and 40 <= humidity <= 75) != (
            2000 <= light <= 8000 and 50 <= moisture <= 70 and 40 <= humidity <= 75):
        triggered.add(22)
    if (2000 <= light <= 6000 and 50 <= moisture <= 70 and 40 <= humidity <= 75) != (
            2000 <= light <= 6000 and 40 <= moisture <= 70 and 40 <= humidity <= 75):
        triggered.add(23)
    if (2000 <= light <= 6000 and 50 <= moisture <= 70 and 40 <= humidity <= 75) != (
            2000 <= light <= 6000 and 50 <= moisture <= 70 and 30 <= humidity <= 75):
        triggered.add(24)
    if (2000 <= light <= 6000 and 50 <= moisture <= 70 and 40 <= humidity <= 75) != (
            2000 <= light <= 6000 and 40 <= moisture <= 70 and 40 <= humidity <= 75):
        triggered.add(25)
    if (2000 <= light <= 6000 and 50 <= moisture <= 70 and 40 <= humidity <= 75) != (
            2000 <= light <= 6000 and 50 <= moisture <= 70 or 40 <= humidity <= 75):
        triggered.add(26)

    if (light > 5500 and moisture > 68 and humidity > 65) != (
            light > 7500 and moisture > 68 and humidity > 65):
        triggered.add(27)
    if (light > 5500 and moisture > 68 and humidity > 65) != (
            light > 8500 and moisture > 68 and humidity > 65):
        triggered.add(28)
    if (light > 5500 and moisture > 68 and humidity > 65) != (
            light < 5500 and moisture > 68 and humidity > 65):
        triggered.add(29)
    if (light > 5500 and moisture > 68 and humidity > 65) != (
            light > 5500 and moisture > 48 and humidity > 65):
        triggered.add(30)
    if (light > 5500 and moisture > 68 and humidity > 65) != (
            light > 5500 and moisture > 38 and humidity > 65):
        triggered.add(31)
    if (light > 5500 and moisture > 68 and humidity > 65) != (
            light > 5500 and moisture > 68 and humidity < 65):
        triggered.add(32)
    if (light > 5500 and moisture > 68 and humidity > 65) != (
            light > 5500 and moisture > 68 and humidity > 35):
        triggered.add(33)
    if (light > 5500 and moisture > 68 and humidity > 65) != (
            light > 5500 and moisture > 68 or humidity > 65):
        triggered.add(34)

    if (light > 7500 and moisture > 28 and humidity > 45) != (
            light > 4500 and moisture > 28 and humidity > 45):
        triggered.add(35)
    if (light > 7500 and moisture > 28 and humidity > 45) != (
            light > 2500 and moisture > 28 and humidity > 45):
        triggered.add(36)
    if (light > 7500 and moisture > 28 and humidity > 45) != (
            light < 7500 and moisture > 28 and humidity > 45):
        triggered.add(37)
    if (light > 7500 and moisture > 28 and humidity > 45) != (
            light > 7500 and moisture > 68 and humidity > 45):
        triggered.add(38)
    if (light > 7500 and moisture > 28 and humidity > 45) != (
            light > 7500 and moisture > 48 and humidity > 45):
        triggered.add(39)
    if (light > 7500 and moisture > 28 and humidity > 45) != (
            light > 7500 and moisture > 28 and humidity < 45):
        triggered.add(40)
    if (light > 7500 and moisture > 28 and humidity > 45) != (
            light > 7500 and moisture > 28 and humidity > 75):
        triggered.add(41)
    if (light > 7500 and moisture > 28 and humidity > 45) != (
            light > 7500 and moisture > 28 and humidity > 65):
        triggered.add(42)

    if (4500 <= light <= 6500 and abs(moisture - humidity) > 25) != (
            2500 <= light <= 6500 and abs(moisture - humidity) > 25):
        triggered.add(43)
    if (4500 <= light <= 6500 and abs(moisture - humidity) > 25) != (
            2500 <= light <= 6500 and abs(moisture - humidity) > 25):
        triggered.add(44)
    if (4500 <= light <= 6500 and abs(moisture - humidity) > 25) != (
            4500 <= light <= 9500 and abs(moisture - humidity) > 25):
        triggered.add(45)
    if (4500 <= light <= 6500 and abs(moisture - humidity) > 25) != (
            4500 <= light <= 8500 and abs(moisture - humidity) > 25):
        triggered.add(46)
    if (4500 <= light <= 6500 and abs(moisture - humidity) > 25) != (
            4500 <= light <= 6500 and abs(moisture + humidity) > 25):
        triggered.add(47)
    if (4500 <= light <= 6500 and abs(moisture - humidity) > 25) != (
            4500 <= light <= 6500 and abs(moisture - humidity + 250) > 25):
        triggered.add(48)
    if (4500 <= light <= 6500 and abs(moisture - humidity) > 25) != (
            4500 <= light <= 6500 and abs(moisture - humidity) > 5):
        triggered.add(49)
    if (4500 <= light <= 6500 and abs(moisture - humidity) > 25) != (
            4500 <= light <= 6500 and abs(moisture - humidity) > 10):
        triggered.add(50)

    if (2500 <= light <= 5500 and 38 <= moisture <= 68 and humidity < 45) != (
            1500 <= light <= 5500 and 38 <= moisture <= 68 and humidity < 45):
        triggered.add(51)
    if (2500 <= light <= 5500 and 38 <= moisture <= 68 and humidity < 45) != (
            2500 <= light <= 7500 and 38 <= moisture <= 68 and humidity < 45):
        triggered.add(52)
    if (2500 <= light <= 5500 and 38 <= moisture <= 68 and humidity < 45) != (
            2500 <= light <= 8500 and 38 <= moisture <= 68 and humidity < 45):
        triggered.add(53)
    if (2500 <= light <= 5500 and 38 <= moisture <= 68 and humidity < 45) != (
            2500 <= light <= 5500 and 48 <= moisture <= 68 and humidity < 45):
        triggered.add(54)
    if (2500 <= light <= 5500 and 38 <= moisture <= 68 and humidity < 45) != (
            2500 <= light <= 5500 and 38 <= moisture <= 85 and humidity < 45):
        triggered.add(55)
    if (2500 <= light <= 5500 and 38 <= moisture <= 68 and humidity < 45) != (
            2500 <= light <= 5500 and 38 <= moisture <= 68 and humidity < 55):
        triggered.add(56)
    if (2500 <= light <= 5500 and 38 <= moisture <= 68 and humidity < 45) != (
            2500 <= light <= 5500 and 38 <= moisture <= 68 and humidity < 75):
        triggered.add(57)
    if (2500 <= light <= 5500 and 38 <= moisture <= 68 and humidity < 45) != (
            2500 <= light <= 5500 and 38 <= moisture <= 68 or humidity < 45):
        triggered.add(58)
    if (2500 <= light <= 5500 and 38 <= moisture <= 68 and humidity < 45) != (
            2500 <= light <= 5500 or 38 <= moisture <= 68 and humidity < 45):
        triggered.add(59)

    if (light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70) != (
            light > 1000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70):
        triggered.add(60)
    if (light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70) != (
            light > 3000 and light < 10000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70):
        triggered.add(61)
    if (light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70) != (
            light > 3000 and light < 7000 or moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70):
        triggered.add(62)
    if (light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70) != (
            light > 3000 and light < 7000 and moisture > 40 and moisture < 75 and humidity > 55 and humidity < 70):
        triggered.add(63)
    if (light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70) != (
            light > 3000 and light < 7000 and moisture > 50 and moisture < 75 and humidity > 55 and humidity < 70):
        triggered.add(64)
    if (light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70) != (
            light > 3000 and light < 7000 and moisture > 60 and moisture < 85 and humidity > 55 and humidity < 70):
        triggered.add(65)
    if (light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70) != (
            light > 3000 and light < 7000 and moisture > 60 or moisture < 75 and humidity > 55 and humidity < 70):
        triggered.add(66)
    if (light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70) != (
            light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 45 and humidity < 70):
        triggered.add(67)
    if (light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70) != (
            light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 35 and humidity < 70):
        triggered.add(68)
    if (light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70) != (
            light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 80):
        triggered.add(69)
    if (light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70) != (
            light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 or humidity < 70):
        triggered.add(70)
    if (light > 3000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70) != (
            light > 1000 and light < 7000 and moisture > 60 and moisture < 75 and humidity > 55 and humidity < 70):
        triggered.add(71)

    if (1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 60) != (
            1800 <= light <= 9200 and 38 <= moisture <= 66 and 46 <= humidity <= 60):
        triggered.add(72)
    if (1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 60) != (
            1800 <= light <= 5900 and 38 <= moisture <= 66 and 46 <= humidity <= 60):
        triggered.add(73)
    if (1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 60) != (
            1800 <= light <= 6200 and 48 <= moisture <= 66 and 46 <= humidity <= 60):
        triggered.add(74)
    if (1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 60) != (
            1800 <= light <= 6200 and 58 <= moisture <= 66 and 46 <= humidity <= 60):
        triggered.add(75)
    if (1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 60) != (
            1800 <= light <= 6200 and 38 <= moisture <= 76 and 46 <= humidity <= 60):
        triggered.add(76)
    if (1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 60) != (
            1800 <= light <= 6200 or 38 <= moisture <= 66 and 46 <= humidity <= 60):
        triggered.add(77)
    if (1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 60) != (
            1800 <= light <= 6200 and 38 <= moisture <= 66 and 36 <= humidity <= 60):
        triggered.add(78)
    if (1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 60) != (
            1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 50):
        triggered.add(79)
    if (1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 60) != (
            1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 70):
        triggered.add(80)
    if (1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 60) != (
            1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 80):
        triggered.add(81)
    if (1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 60) != (
            1800 <= light <= 6200 and 38 <= moisture <= 66 or 46 <= humidity <= 60):
        triggered.add(82)
    if (1800 <= light <= 6200 and 38 <= moisture <= 66 and 46 <= humidity <= 60) != (
            1800 <= light <= 8200 and 38 <= moisture <= 66 and 46 <= humidity <= 60):
        triggered.add(83)

    if (light > 6500 and moisture > 70 and humidity < 48) != (
            light > 5500 and moisture > 70 and humidity < 48):
        triggered.add(84)
    if (light > 6500 and moisture > 70 and humidity < 48) != (
            light > 4500 and moisture > 70 and humidity < 48):
        triggered.add(85)
    if (light > 6500 and moisture > 70 and humidity < 48) != (
            light > 8500 and moisture > 70 and humidity < 48):
        triggered.add(86)
    if (light > 6500 and moisture > 70 and humidity < 48) != (
            light > 6500 and moisture > 40 and humidity < 48):
        triggered.add(87)
    if (light > 6500 and moisture > 70 and humidity < 48) != (
            light > 6500 and moisture > 20 and humidity < 48):
        triggered.add(88)
    if (light > 6500 and moisture > 70 and humidity < 48) != (
            light > 6500 and moisture > 70 and humidity < 38):
        triggered.add(89)
    if (light > 6500 and moisture > 70 and humidity < 48) != (
            light > 6500 and moisture > 70 and humidity < 28):
        triggered.add(90)
    if (light > 6500 and moisture > 70 and humidity < 48) != (
            light > 6500 or moisture > 70 and humidity < 48):
        triggered.add(91)
    if (light > 6500 and moisture > 70 and humidity < 48) != (
            light > 6500 and moisture > 70 or humidity < 48):
        triggered.add(92)
    if (light > 6500 and moisture > 70 and humidity < 48) != (
            light > 6500 and moisture > 70 and humidity > 48):
        triggered.add(93)

    if (light < 6800 and moisture < 42) != (light < 4800 and moisture < 42):
        triggered.add(94)
    if (light < 6800 and moisture < 42) != (light < 9800 and moisture < 42):
        triggered.add(95)
    if (light < 6800 and moisture < 42) != (light > 3800 and moisture < 42):
        triggered.add(96)
    if (light < 6800 and moisture < 42) != (light < 6800 or moisture < 42):
        triggered.add(97)
    if (light < 6800 and moisture < 42) != (light < 6800 and moisture < 62):
        triggered.add(98)
    if (light < 6800 and moisture < 42) != (light < 6800 and moisture < 72):
        triggered.add(99)
    if (light < 6800 and moisture < 42) != (light < 6800 and moisture < 82):
        triggered.add(100)

    return triggered

# ===  ===
# === 状态处理辅助函数 ===
def clip_state(state):
    return np.clip(state, EXPERIMENT_CONFIG['MIN_VALUES'], EXPERIMENT_CONFIG['MAX_VALUES'])

def denormalize_state(normalized_state):
    """将归一化状态还原为原始状态"""
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return normalized_state * (max_vals - min_vals) / 2 + (min_vals + max_vals) / 2

def coverage_similarity(triggered, target_path):
    """
    Similarity:  / target paths
    
    """
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0

    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)

def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']

    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']

    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']

    return reward

# === PPO Actor ===
class PPOActor(nn.Module):
    def __init__(self, state_dim=3, action_dim=3, hidden_dim=256):
        super(PPOActor, self).__init__()

        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)

        self.mean_head = nn.Linear(hidden_dim, action_dim)
        self.log_std_head = nn.Linear(hidden_dim, action_dim)

        self.action_scale = 10.0

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))
        x = F.relu(self.fc3(x))

        mean = torch.tanh(self.mean_head(x)) * self.action_scale
        log_std = torch.clamp(self.log_std_head(x), -20, 2)

        return mean, log_std

    def get_action_and_log_prob(self, state):
        mean, log_std = self.forward(state)
        std = torch.exp(log_std)

        dist = torch.distributions.Normal(mean, std)
        action = dist.sample()
        log_prob = dist.log_prob(action).sum(dim=-1)

        return action, log_prob

    def get_log_prob(self, state, action):
        mean, log_std = self.forward(state)
        std = torch.exp(log_std)

        dist = torch.distributions.Normal(mean, std)
        log_prob = dist.log_prob(action).sum(dim=-1)

        return log_prob

# === PPO Critic ===
class PPOCritic(nn.Module):
    def __init__(self, state_dim=3, hidden_dim=256):
        super(PPOCritic, self).__init__()

        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)
        self.value_head = nn.Linear(hidden_dim, 1)

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))
        x = F.relu(self.fc3(x))
        value = self.value_head(x)
        return value.squeeze(-1)

# === PPO ===
class PPOBuffer:
    def __init__(self):
        self.states = []
        self.actions = []
        self.rewards = []
        self.values = []
        self.log_probs = []
        self.dones = []
        self.path_indices = []
        self.similarities = []

    def store(self, state, action, reward, value, log_prob, done, path_idx, similarity):
        self.states.append(state)
        self.actions.append(action)
        self.rewards.append(reward)
        self.values.append(value)
        self.log_probs.append(log_prob)
        self.dones.append(done)
        self.path_indices.append(path_idx)
        self.similarities.append(similarity)

    def compute_advantages(self):
        rewards = np.array(self.rewards)
        values = np.array(self.values)
        dones = np.array(self.dones)

        advantages = np.zeros_like(rewards)
        returns = np.zeros_like(rewards)
        last_advantage = 0

        for t in reversed(range(len(rewards))):
            if t == len(rewards) - 1:
                next_value = 0
            else:
                next_value = values[t + 1]

            delta = rewards[t] + EXPERIMENT_CONFIG['GAMMA'] * next_value * (1 - dones[t]) - values[t]
            advantages[t] = delta + EXPERIMENT_CONFIG['GAMMA'] * EXPERIMENT_CONFIG['GAE_LAMBDA'] * (
                    1 - dones[t]) * last_advantage
            last_advantage = advantages[t]

        returns = advantages + values
        advantages = (advantages - np.mean(advantages)) / (np.std(advantages) + 1e-8)

        return advantages, returns

    def get_batch_iterator(self, advantages, returns):
        indices = np.arange(len(self.states))
        np.random.shuffle(indices)

        batch_size = EXPERIMENT_CONFIG['PPO_BATCH_SIZE']
        for start in range(0, len(self.states), batch_size):
            end = min(start + batch_size, len(self.states))
            batch_indices = indices[start:end]

            yield {
                'states': torch.FloatTensor(np.array([self.states[i] for i in batch_indices])).to(device),
                'actions': torch.FloatTensor(np.array([self.actions[i] for i in batch_indices])).to(device),
                'old_log_probs': torch.FloatTensor(np.array([self.log_probs[i] for i in batch_indices])).to(device),
                'advantages': torch.FloatTensor(advantages[batch_indices]).to(device),
                'returns': torch.FloatTensor(returns[batch_indices]).to(device)
            }

    def get_top_k_per_path(self, num_paths, k=20):
        path_samples = {i: [] for i in range(num_paths)}

        for idx in range(len(self.states)):
            path_idx = self.path_indices[idx]
            similarity = self.similarities[idx]
            state = self.states[idx]
            path_samples[path_idx].append((idx, similarity, state))

        top_k_results = {}
        for path_idx in range(num_paths):
            samples = path_samples[path_idx]
            if len(samples) == 0:
                top_k_results[path_idx] = []
                continue

            samples.sort(key=lambda x: x[1], reverse=True)
            top_k = samples[:k]

            top_k_results[path_idx] = []
            for sample in top_k:
                normalized_state = sample[2]
                original_state = denormalize_state(normalized_state)
                original_state_int = np.round(original_state).astype(int)

                top_k_results[path_idx].append({
                    'state': original_state_int,
                    'similarity': sample[1],
                    'triggered': execute_Tr(*original_state_int)
                })

        return top_k_results

    def clear(self):
        self.states.clear()
        self.actions.clear()
        self.rewards.clear()
        self.values.clear()
        self.log_probs.clear()
        self.dones.clear()
        self.path_indices.clear()
        self.similarities.clear()

    def __len__(self):
        return len(self.states)

# === PPO ===
class PPOAgent:
    def __init__(self, state_dim=3, action_dim=3):
        self.state_dim = state_dim
        self.action_dim = action_dim

        self.actor = PPOActor(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.actor_optimizer = optim.Adam(self.actor.parameters(), lr=EXPERIMENT_CONFIG['ACTOR_LR'])

        self.critic = PPOCritic(state_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_optimizer = optim.Adam(self.critic.parameters(), lr=EXPERIMENT_CONFIG['CRITIC_LR'])

        self.buffer = PPOBuffer()
        self.update_count = 0

    def get_action(self, state):
            min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
            max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
            normalized_state = (state - (min_vals + max_vals) / 2) / ((max_vals - min_vals) / 2)
            state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)

            with torch.no_grad():
                action, log_prob = self.actor.get_action_and_log_prob(state_tensor)
                value = self.critic(state_tensor)

            action = action.cpu().numpy()[0]
            log_prob = log_prob.cpu().item()
            value = value.cpu().item()

            return action, log_prob, value
    def store_experience(self, state, action, reward, value, log_prob, done, path_idx, similarity):
            min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
            max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
            normalized_state = (state - (min_vals + max_vals) / 2) / ((max_vals - min_vals) / 2)
            self.buffer.store(normalized_state, action, reward, value, log_prob, done, path_idx, similarity)

    def update(self):
        if len(self.buffer) == 0:
            return

        advantages, returns = self.buffer.compute_advantages()

        for epoch in range(EXPERIMENT_CONFIG['PPO_EPOCHS']):
            for batch in self.buffer.get_batch_iterator(advantages, returns):
                new_log_probs = self.actor.get_log_prob(batch['states'], batch['actions'])
                ratio = torch.exp(new_log_probs - batch['old_log_probs'])

                surr1 = ratio * batch['advantages']
                surr2 = torch.clamp(ratio, 1 - EXPERIMENT_CONFIG['CLIP_EPSILON'],
                                    1 + EXPERIMENT_CONFIG['CLIP_EPSILON']) * batch['advantages']
                actor_loss = -torch.min(surr1, surr2).mean()

                self.actor_optimizer.zero_grad()
                actor_loss.backward()
                torch.nn.utils.clip_grad_norm_(self.actor.parameters(), 0.5)
                self.actor_optimizer.step()

                new_values = self.critic(batch['states'])
                critic_loss = F.mse_loss(new_values, batch['returns'])

                self.critic_optimizer.zero_grad()
                critic_loss.backward()
                torch.nn.utils.clip_grad_norm_(self.critic.parameters(), 0.5)
                self.critic_optimizer.step()

        self.update_count += 1
        self.buffer.clear()

        if self.update_count % 2 == 0:
            print(f"  -> PPOcompleted (Run {self.update_count})")

# === Metric ===
def calculate_run_performance(run_idx, ppo_results, training_time, total_steps, update_count, agent):
    """ runMetric"""
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    # 1. (Total Reward)
    total_reward = 0
    # 2. (Average Reward)
    average_reward = 0
    # 5. (Convergence)
    convergence = 0
    # 12. (Environment Adaptability)
    environment_adaptability = 0
    # 13. (Generalization Ability)
    generalization_ability = 0
    # 15. (Computational Efficiency)
    computational_efficiency = 0
    # 16. (Policy Update Frequency)
    policy_update_frequency = 0

    # Similarity
    all_similarities = []

    # Metric
    total_samples = 0
    all_rewards = []

    for path_idx in range(num_paths):
        samples = ppo_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']

            total_reward += reward
            all_rewards.append(reward)
            all_similarities.append(similarity)
            total_samples += 1

    # 1. 
    total_reward = total_reward

    # 2. 
    if total_samples > 0:
        average_reward = total_reward / total_samples

    # 5. (Average Similarity)
    if all_similarities:
        convergence = np.mean(all_similarities)

    # 12. (Similarity)
    if len(all_similarities) > 1:
        environment_adaptability = 1 / (np.std(all_similarities) + 1e-8)

    # 13. (Average Similarity)
    generalization_ability = convergence

    # 15. (/ seconds)
    if training_time > 0:
        computational_efficiency = total_steps / training_time

    # 16. 
    if training_time > 0:
        policy_update_frequency = update_count / training_time

    # Similarity
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0

    return {
        'Run': run_idx + 1,

        # Metric
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),

        # Similarity
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }

# === Excel ===
def export_to_excel(all_ppo_results, all_performance_data, target_paths, output_path="PPO_20 run.xlsx"):
    """20 runPPOExcel"""
    print("\nExcel...")

    # 
    all_ppo_summary_data = []
    all_ppo_detailed_data = []

    #  run
    for run_idx, (ppo_results, performance_data) in enumerate(zip(all_ppo_results, all_performance_data)):
        # ===== Sheet1: PPOPath  =====
        ppo_summary_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = ppo_results[path_idx]

            if len(samples) == 0:
                ppo_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue

            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = '' if perfect_count > 0 else ''

            ppo_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })

        all_ppo_summary_data.extend(ppo_summary_data)

        # ===== Sheet2: PPODetailed Sample Data =====
        ppo_detailed_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = ppo_results[path_idx]

            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']

                ppo_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })

        all_ppo_detailed_data.extend(ppo_detailed_data)

    # Excel
    ppo_summary_df = pd.DataFrame(all_ppo_summary_data)
    ppo_detailed_df = pd.DataFrame(all_ppo_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet1: PPOPath 
        ppo_summary_df.to_excel(writer, sheet_name='PPOPath ', index=False)

        # Sheet2: PPODetailed Sample Data
        ppo_detailed_df.to_excel(writer, sheet_name='PPODetailed Sample Data', index=False)

        # Sheet3: Metric - 
        selected_columns = [
            'Run',
            'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df_selected = performance_df[selected_columns]
        performance_df_selected.to_excel(writer, sheet_name='Metric', index=False)

        # 
        workbook = writer.book

        # 
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        perfect_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # 

        # === Sheet1 ===
        ws1 = writer.sheets['PPOPath ']
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 
        for row_idx in range(2, ws1.max_row + 1):
            if ws1.cell(row_idx, 9).value == '':  # Run 9""
                for col_idx in range(1, ws1.max_column + 1):
                    ws1.cell(row_idx, col_idx).fill = perfect_fill

        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 15
        ws1.column_dimensions['F'].width = 15
        ws1.column_dimensions['G'].width = 15
        ws1.column_dimensions['H'].width = 15
        ws1.column_dimensions['I'].width = 15
        ws1.column_dimensions['J'].width = 50

        # === Sheet2 ===
        ws2 = writer.sheets['PPODetailed Sample Data']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 15
        ws2.column_dimensions['I'].width = 40
        ws2.column_dimensions['J'].width = 40
        ws2.column_dimensions['K'].width = 15
        ws2.column_dimensions['L'].width = 15

        # === Sheet3 ===
        ws3 = writer.sheets['Metric']
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        for col in columns:
            ws3.column_dimensions[col].width = 18

    print(f"Excel: {output_path}")
    print(f"  - Sheet1: PPOPath  ({len(all_ppo_summary_data)})")
    print(f"  - Sheet2: PPODetailed Sample Data ({len(all_ppo_detailed_data)})")
    print(f"  - Sheet3: Metric ({len(all_performance_data)})")

# === PPO ===
def train_ppo_workflow():
    print("=" * 80)
    print("PPO")
    print("Similarity:  / target paths")
    print("=" * 80)

    agent = PPOAgent()
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n: Path {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']}")
    path_samples = {}
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']

    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            # 分别生成 X, Y, Z 的随机整数
            state = np.array([
                np.random.randint(min_vals[0], max_vals[0] + 1),
                np.random.randint(min_vals[1], max_vals[1] + 1),
                np.random.randint(min_vals[2], max_vals[2] + 1)
            ], dtype=np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  Path  {path_idx + 1}/{num_paths}:  {len(samples)} ")

    # 
    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size

    print(f"\n: {batch_size},{EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(f": {num_batches} /Path  x {num_paths} Path  = {num_batches * num_paths} ")
    print("-" * 80)

    global_buffer = PPOBuffer()

    for batch_idx in range(num_batches):
        print(f"\n {batch_idx + 1}/{num_batches}")

        for path_idx in range(num_paths):
            target_path = target_paths[path_idx]
            batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]

            batch_rewards = []
            batch_similarities = []

            for sample_idx, initial_state in enumerate(batch_samples):
                state = initial_state.copy()
                episode_reward = 0
                final_similarity = 0

                for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                    action, log_prob, value = agent.get_action(state)

                    next_state = state + action
                    next_state = clip_state(next_state)

                    triggered = execute_Tr(*next_state)
                    reward = unified_reward_function(triggered, target_path)
                    similarity = coverage_similarity(triggered, target_path)

                    done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)

                    agent.store_experience(state, action, reward, value, log_prob, done, path_idx, similarity)
                    normalized_state = (state - (min_vals + max_vals) / 2) / ((max_vals - min_vals) / 2)
                    global_buffer.store(
                        normalized_state,
                        action, reward, value, log_prob, done, path_idx, similarity
                    )

                    state = next_state
                    episode_reward += reward
                    final_similarity = similarity
                    total_steps += 1

                batch_rewards.append(episode_reward)
                batch_similarities.append(final_similarity)

            avg_reward = np.mean(batch_rewards)
            avg_similarity = np.mean(batch_similarities)
            print(f"  Path {path_idx + 1}: ={avg_reward:.2f}, Average Similarity={avg_similarity:.4f}")

        print(f"\n  PPO...")
        agent.update()
        print(f"  : {len(global_buffer)}")

    training_time = time.time() - start_time

    print("\n" + "=" * 80)
    print(f"PPOcompleted! Total elapsed time: {training_time:.2f} seconds, : {total_steps}")
    print(f": {len(global_buffer)}")
    print(f"PPO: {agent.update_count}")
    print("=" * 80)

    # Top-K
    print(f"\nPath SimilarityMaximum{EXPERIMENT_CONFIG['TOP_K_SAMPLES']}...")
    ppo_top_k_results = global_buffer.get_top_k_per_path(num_paths, EXPERIMENT_CONFIG['TOP_K_SAMPLES'])

    return agent, ppo_top_k_results, training_time, total_steps, agent.update_count

# ===  ===
def main():
    print("\n" + "=" * 80)
    print("PPO - 20 run")
    print("Metric")
    print("=" * 80)

    all_ppo_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    # 20
    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'=' * 80}")
        print(f"Start run  {run_idx + 1}/{EXPERIMENT_CONFIG['NUM_RUNS']}  run")
        print(f"{'=' * 80}")

        # PPO
        ppo_agent, ppo_results, training_time, total_steps, update_count = train_ppo_workflow()

        # Metric
        performance_data = calculate_run_performance(
            run_idx, ppo_results, training_time, total_steps, update_count, ppo_agent
        )

        # 
        all_ppo_results.append(ppo_results)
        all_performance_data.append(performance_data)

        print(f"\nRun  {run_idx + 1}  runcompleted!")
        print(f"  Total Reward: {performance_data['Total Reward']}")
        print(f"  Average Reward: {performance_data['Average Reward']}")
        print(f"  Average Similarity: {performance_data['Average Similarity']}")

    # Excel(20 run)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"PPO_20 run_{timestamp}.xlsx"
    export_to_excel(all_ppo_results, all_performance_data, target_paths, output_path)

    # 
    print("\n" + "=" * 80)
    print("20 run")
    print("=" * 80)

    # Metric
    total_rewards = [p['Total Reward'] for p in all_performance_data]
    average_rewards = [p['Average Reward'] for p in all_performance_data]
    convergences = [p['Convergence'] for p in all_performance_data]
    environment_adaptabilities = [p['Environment Adaptability'] for p in all_performance_data]
    generalization_abilities = [p['Generalization Ability'] for p in all_performance_data]
    computational_efficiencies = [p['Computational Efficiency'] for p in all_performance_data]
    policy_update_frequencies = [p['Policy Update Frequency'] for p in all_performance_data]
    avg_similarities = [p['Average Similarity'] for p in all_performance_data]

    print(f"\nTotal Reward Statistics:")
    print(f"  Mean: {np.mean(total_rewards):.2f}")
    print(f"  Standard deviation: {np.std(total_rewards):.2f}")

    print(f"\nAverage Reward Statistics:")
    print(f"  Mean: {np.mean(average_rewards):.4f}")
    print(f"  Standard deviation: {np.std(average_rewards):.4f}")

    print(f"\nConvergence Statistics:")
    print(f"  Mean: {np.mean(convergences):.4f}")
    print(f"  Standard deviation: {np.std(convergences):.4f}")

    print(f"\nEnvironment Adaptability Statistics:")
    print(f"  Mean: {np.mean(environment_adaptabilities):.4f}")
    print(f"  Standard deviation: {np.std(environment_adaptabilities):.4f}")

    print(f"\nGeneralization Ability Statistics:")
    print(f"  Mean: {np.mean(generalization_abilities):.4f}")
    print(f"  Standard deviation: {np.std(generalization_abilities):.4f}")

    print(f"\nComputational Efficiency Statistics:")
    print(f"  Mean: {np.mean(computational_efficiencies):.2f}")
    print(f"  Standard deviation: {np.std(computational_efficiencies):.2f}")

    print(f"\nPolicy Update Frequency Statistics:")
    print(f"  Mean: {np.mean(policy_update_frequencies):.4f}")
    print(f"  Standard deviation: {np.std(policy_update_frequencies):.4f}")

    print("\n" + "=" * 80)
    print(f" {EXPERIMENT_CONFIG['NUM_RUNS']} completed!")
    print("=" * 80)

if __name__ == "__main__":
    main()