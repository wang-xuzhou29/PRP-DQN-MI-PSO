import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

def execute_validation_rules(light, moisture, temp):
    actions = []
    triggered = set()

    if (light < 3500 and moisture < 55) != (light < 5500 and moisture < 55): triggered.add(1)
    if (light < 3500 and moisture < 55) != (light < 6500 and moisture < 55): triggered.add(2)
    if (light < 3500 and moisture < 55) != (light < 3500 and moisture < 45): triggered.add(3)
    if (light < 3500 and moisture < 55) != (light < 3500 and moisture < 35): triggered.add(4)
    if (light < 3500 and moisture < 55) != (light < 3500 and moisture < 75): triggered.add(5)

    if (light < 2200 and moisture < 48 and temp < 20) != (light < 4200 and moisture < 48 and temp < 20): triggered.add(6)
    if (light < 2200 and moisture < 48 and temp < 20) != (light < 6200 and moisture < 48 and temp < 20): triggered.add(7)
    if (light < 2200 and moisture < 48 and temp < 20) != (light < 2200 and moisture < 58 and temp < 20): triggered.add(8)
    if (light < 2200 and moisture < 48 and temp < 20) != (light < 2200 and moisture < 48 and temp < 30): triggered.add(9)
    if (light < 2200 and moisture < 48 and temp < 20) != (light < 2200 and moisture < 48 and temp > 20): triggered.add(10)
    if (light < 2200 and moisture < 48 and temp < 20) != (light < 2200 and moisture > 48 and temp < 20): triggered.add(11)
    if (light < 2200 and moisture < 48 and temp < 20) != (light > 2200 and moisture < 48 and temp < 20): triggered.add(12)

    if ((light < 1800 or temp < 18) and moisture < 45) != ((light < 1800 and temp < 18) and moisture < 45): triggered.add(13)
    if ((light < 1800 or temp < 18) and moisture < 45) != ((light < 1800 or temp < 18) or moisture < 45): triggered.add(14)
    if ((light < 1800 or temp < 18) and moisture < 45) != ((light > 1800 or temp < 18) and moisture < 45): triggered.add(15)
    if ((light < 1800 or temp < 18) and moisture < 45) != ((light < 1800 or temp > 18) and moisture < 45): triggered.add(16)
    if ((light < 1800 or temp < 18) and moisture < 45) != ((light < 1800 or temp < 18) and moisture > 45): triggered.add(17)
    if ((light < 1800 or temp < 18) and moisture < 45) != ((light < 1800 or temp < 38) and moisture < 45): triggered.add(18)
    if ((light < 1800 or temp < 18) and moisture < 45) != ((light < 1800 or temp < 18) and moisture < 65): triggered.add(19)

    if (light * moisture < 160000) != (light * moisture < 130000): triggered.add(20)
    if (light * moisture < 160000) != (light * moisture < 60000): triggered.add(21)
    if (light * moisture < 160000) != (light * moisture < 90000): triggered.add(22)

    if (light > 2500 and light < 3500 and moisture > 50 and moisture < 58) != (light > 1500 and light < 3500 and moisture > 50 and moisture < 58): triggered.add(23)
    if (light > 2500 and light < 3500 and moisture > 50 and moisture < 58) != (light > 500 and light < 3500 and moisture > 50 and moisture < 58): triggered.add(24)
    if (light > 2500 and light < 3500 and moisture > 50 and moisture < 58) != (light < 2500 and light < 3500 and moisture > 50 and moisture < 58): triggered.add(25)
    if (light > 2500 and light < 3500 and moisture > 50 and moisture < 58) != (light > 2500 and light > 3500 and moisture > 50 and moisture < 58): triggered.add(26)
    if (light > 2500 and light < 3500 and moisture > 50 and moisture < 58) != (light > 2500 and light < 6500 and moisture > 50 and moisture < 58): triggered.add(27)
    if (light > 2500 and light < 3500 and moisture > 50 and moisture < 58) != (light > 2500 and light < 8500 and moisture > 50 and moisture < 58): triggered.add(28)
    if (light > 2500 and light < 3500 and moisture > 50 and moisture < 58) != (light > 2500 and light < 3500 and moisture < 50 and moisture < 58): triggered.add(29)
    if (light > 2500 and light < 3500 and moisture > 50 and moisture < 58) != (light > 2500 and light < 3500 and moisture > 50 and moisture > 58): triggered.add(30)
    if (light > 2500 and light < 3500 and moisture > 50 and moisture < 58) != (light > 2500 and light < 3500 and moisture > 50 and moisture < 38): triggered.add(31)
    if (light > 2500 and light < 3500 and moisture > 50 and moisture < 58) != (light > 2500 and light < 3500 and moisture > 20 and moisture < 58): triggered.add(32)
    if (light > 2500 and light < 3500 and moisture > 50 and moisture < 58) != (light > 2500 and light < 3500 and moisture > 50 or moisture < 58): triggered.add(33)

    if ((3000 - light) > (58 - moisture) * 25) != ((4000 - light) > (58 - moisture) * 25): triggered.add(34)
    if ((3000 - light) > (58 - moisture) * 25) != ((5000 - light) > (58 - moisture) * 25): triggered.add(35)
    if ((3000 - light) > (58 - moisture) * 25) != ((6000 - light) > (58 - moisture) * 25): triggered.add(36)
    if ((3000 - light) > (58 - moisture) * 25) != ((3000 + light) > (58 - moisture) * 25): triggered.add(37)
    if ((3000 - light) > (58 - moisture) * 25) != ((3000 - light) < (58 - moisture) * 25): triggered.add(38)
    if ((3000 - light) > (58 - moisture) * 25) != ((3000 - light) > (68 - moisture) * 25): triggered.add(39)
    if ((3000 - light) > (58 - moisture) * 25) != ((3000 - light) > (78 - moisture) * 25): triggered.add(40)
    if ((3000 - light) > (58 - moisture) * 25) != ((3000 - light) > (58 + moisture) * 25): triggered.add(41)
    if ((3000 - light) > (58 - moisture) * 25) != ((3000 - light) > (58 - moisture) * 15): triggered.add(42)
    if ((3000 - light) > (58 - moisture) * 25) != ((3000 - light) > (58 - moisture) * 35): triggered.add(43)
    if ((3000 - light) > (58 - moisture) * 25) != ((3000 - light) > (58 - moisture) * 45): triggered.add(44)

    if (light > 8500 and temp > 30) != (light > 5500 and temp > 30): triggered.add(45)
    if (light > 8500 and temp > 30) != (light > 3500 and temp > 30): triggered.add(46)
    if (light > 8500 and temp > 30) != (light > 8500 and temp > 20): triggered.add(47)
    if (light > 8500 and temp > 30) != (light < 8500 and temp > 30): triggered.add(48)
    if (light > 8500 and temp > 30) != (light > 8500 and temp < 30): triggered.add(49)
    if (light > 8500 and temp > 30) != (light > 8500 and moisture > 30): triggered.add(50)
    if (light > 8500 and temp > 30) != (light > 8500 and temp + moisture > 30): triggered.add(51)
    if (light > 8500 and temp > 30) != (light > 8500 and moisture - temp > 30): triggered.add(52)
    if (light > 8500 and temp > 30) != (light > 8500 and temp > 15): triggered.add(53)

    if (light > 2000 and light < 4000 and moisture > 40 and moisture < 65) != (light > 1000 and light < 4000 and moisture > 40 and moisture < 65): triggered.add(54)
    if (light > 2000 and light < 4000 and moisture > 40 and moisture < 65) != (light > 3000 and light < 4000 and moisture > 40 and moisture < 65): triggered.add(55)
    if (light > 2000 and light < 4000 and moisture > 40 and moisture < 65) != (light > 3500 and light < 4000 and moisture > 40 and moisture < 65): triggered.add(56)
    if (light > 2000 and light < 4000 and moisture > 40 and moisture < 65) != (light > 1500 and light < 4000 and moisture > 40 and moisture < 65): triggered.add(57)
    if (light > 2000 and light < 4000 and moisture > 40 and moisture < 65) != (light < 2000 and light < 4000 and moisture > 40 and moisture < 65): triggered.add(58)
    if (light > 2000 and light < 4000 and moisture > 40 and moisture < 65) != (light > 2000 and light > 4000 and moisture > 40 and moisture < 65): triggered.add(59)
    if (light > 2000 and light < 4000 and moisture > 40 and moisture < 65) != (light > 2000 and light < 4000 or moisture > 40 and moisture < 65): triggered.add(60)
    if (light > 2000 and light < 4000 and moisture > 40 and moisture < 65) != (light > 2000 and light < 4000 and moisture > 50 and moisture < 65): triggered.add(61)
    if (light > 2000 and light < 4000 and moisture > 40 and moisture < 65) != (light > 2000 and light < 4000 and moisture > 40 and moisture > 65): triggered.add(62)
    if (light > 2000 and light < 4000 and moisture > 40 and moisture < 65) != (light > 2000 and light < 4000 and moisture < 40 and moisture < 65): triggered.add(63)

    if (moisture < 50 and (light < 2000 or temp > 30)) != (moisture > 50 and (light < 2000 or temp > 30)): triggered.add(64)
    if (moisture < 50 and (light < 2000 or temp > 30)) != (moisture < 20 and (light < 2000 or temp > 30)): triggered.add(65)
    if (moisture < 50 and (light < 2000 or temp > 30)) != (moisture < 50 and (light > 2000 or temp > 30)): triggered.add(66)
    if (moisture < 50 and (light < 2000 or temp > 30)) != (moisture < 50 and (light < 3000 or temp > 30)): triggered.add(67)
    if (moisture < 50 and (light < 2000 or temp > 30)) != (moisture < 50 and (light < 4000 or temp > 30)): triggered.add(68)
    if (moisture < 50 and (light < 2000 or temp > 30)) != (moisture < 50 and (light < 2000 and temp > 30)): triggered.add(69)
    if (moisture < 50 and (light < 2000 or temp > 30)) != (moisture < 50 and (light < 2000 or temp < 30)): triggered.add(70)
    if (moisture < 50 and (light < 2000 or temp > 30)) != (moisture < 50 and (light < 2000 or temp > 50)): triggered.add(71)
    if (moisture < 50 and (light < 2000 or temp > 30)) != (moisture < 70 and (light < 2000 or temp > 30)): triggered.add(72)
    if (moisture < 50 and (light < 2000 or temp > 30)) != (moisture < 40 and (light < 2000 or temp > 30)): triggered.add(73)
    if (moisture < 50 and (light < 2000 or temp > 30)) != (moisture < 25 and (light < 2000 or temp > 30)): triggered.add(74)

    if (moisture > 35 and moisture < 50 and light > 1500 and light < 3000) != (moisture > 45 and moisture < 50 and light > 1500 and light < 3000): triggered.add(75)
    if (moisture > 35 and moisture < 50 and light > 1500 and light < 3000) != (moisture > 35 and moisture < 60 and light > 1500 and light < 3000): triggered.add(76)
    if (moisture > 35 and moisture < 50 and light > 1500 and light < 3000) != (moisture > 35 and moisture < 70 and light > 1500 and light < 3000): triggered.add(77)
    if (moisture > 35 and moisture < 50 and light > 1500 and light < 3000) != (moisture < 35 and moisture < 50 and light > 1500 and light < 3000): triggered.add(78)
    if (moisture > 35 and moisture < 50 and light > 1500 and light < 3000) != (moisture > 35 and moisture > 50 and light > 1500 and light < 3000): triggered.add(79)
    if (moisture > 35 and moisture < 50 and light > 1500 and light < 3000) != (moisture > 35 and moisture < 50 and light < 1500 and light < 3000): triggered.add(80)
    if (moisture > 35 and moisture < 50 and light > 1500 and light < 3000) != (moisture > 35 and moisture < 50 and light > 2500 and light < 3000): triggered.add(81)
    if (moisture > 35 and moisture < 50 and light > 1500 and light < 3000) != (moisture > 35 and moisture < 50 and light > 1500 and light > 3000): triggered.add(82)
    if (moisture > 35 and moisture < 50 and light > 1500 and light < 3000) != (moisture > 35 and moisture < 50 and light > 1500 and light < 5000): triggered.add(83)

    if (moisture < 40 + (light - 2000) / 100) != (moisture < 40 + (light - 3000) / 100): triggered.add(84)
    if (moisture < 40 + (light - 2000) / 100) != (moisture < 40 + (light - 1000) / 100): triggered.add(85)
    if (moisture < 40 + (light - 2000) / 100) != (moisture < 50 + (light - 2000) / 100): triggered.add(86)
    if (moisture < 40 + (light - 2000) / 100) != (moisture < 60 + (light - 2000) / 100): triggered.add(87)
    if (moisture < 40 + (light - 2000) / 100) != (moisture < 40 + (light + 2000) / 100): triggered.add(88)
    if (moisture < 40 + (light - 2000) / 100) != (moisture < 40 + (light - 2000) / 200): triggered.add(89)
    if (moisture < 40 + (light - 2000) / 100) != (moisture < 40 + (light - 2000) / 50): triggered.add(90)
    if (moisture < 40 + (light - 2000) / 100) != (moisture > 40 + (light - 2000) / 100): triggered.add(91)
    if (moisture < 40 + (light - 2000) / 100) != (moisture < 60 + (light - 2000) / 100): triggered.add(92)
    if (moisture < 40 + (light - 2000) / 100) != (moisture < 30 + (light - 2000) / 100): triggered.add(93)

    if (moisture < 48 and light < 2800 and temp > 24) != (moisture < 38 and light < 2800 and temp > 24): triggered.add(94)
    if (moisture < 48 and light < 2800 and temp > 24) != (moisture < 28 and light < 2800 and temp > 24): triggered.add(95)
    if (moisture < 48 and light < 2800 and temp > 24) != (moisture < 58 and light < 2800 and temp > 24): triggered.add(96)
    if (moisture < 48 and light < 2800 and temp > 24) != (moisture < 68 and light < 2800 and temp > 24): triggered.add(97)
    if (moisture < 48 and light < 2800 and temp > 24) != (moisture < 48 and light < 3800 and temp > 24): triggered.add(98)
    if (moisture < 48 and light < 2800 and temp > 24) != (moisture < 48 and light < 5800 and temp > 24): triggered.add(99)
    if (moisture < 48 and light < 2800 and temp > 24) != (moisture < 48 and light < 2800 and temp > 34): triggered.add(100)
    if (moisture < 48 and light < 2800 and temp > 24) != (moisture < 48 and light < 2800 and temp > 14): triggered.add(101)
    if (moisture < 48 and light < 2800 and temp > 24) != (moisture > 48 and light < 2800 and temp > 24): triggered.add(102)
    if (moisture < 48 and light < 2800 and temp > 24) != (moisture < 48 and light > 2800 and temp > 24): triggered.add(103)
    if (moisture < 48 and light < 2800 and temp > 24) != (moisture < 48 and light < 2800 and temp < 24): triggered.add(104)

    if (light < 2500 and moisture > 50 and moisture < 60) != (light < 3500 and moisture > 50 and moisture < 60): triggered.add(105)
    if (light < 2500 and moisture > 50 and moisture < 60) != (light < 4500 and moisture > 50 and moisture < 60): triggered.add(106)
    if (light < 2500 and moisture > 50 and moisture < 60) != (light < 6500 and moisture > 50 and moisture < 60): triggered.add(107)
    if (light < 2500 and moisture > 50 and moisture < 60) != (light < 2500 and moisture > 30 and moisture < 60): triggered.add(108)
    if (light < 2500 and moisture > 50 and moisture < 60) != (light < 2500 and moisture > 20 and moisture < 60): triggered.add(109)
    if (light < 2500 and moisture > 50 and moisture < 60) != (light < 2500 and moisture > 50 and moisture < 70): triggered.add(110)
    if (light < 2500 and moisture > 50 and moisture < 60) != (light < 2500 or moisture > 50 and moisture < 60): triggered.add(111)
    if (light < 2500 and moisture > 50 and moisture < 60) != (light < 2500 and moisture > 50 or moisture < 60): triggered.add(112)
    if (light < 2500 and moisture > 50 and moisture < 60) != (light < 2000 and moisture > 50 and moisture < 60): triggered.add(113)

    if (temp > 25 and light < 2500 and moisture < 48) != (temp > 35 and light < 2500 and moisture < 48): triggered.add(114)
    if (temp > 25 and light < 2500 and moisture < 48) != (temp > 15 and light < 2500 and moisture < 48): triggered.add(115)
    if (temp > 25 and light < 2500 and moisture < 48) != (temp > 25 and light < 4500 and moisture < 48): triggered.add(116)
    if (temp > 25 and light < 2500 and moisture < 48) != (temp > 25 and light < 6500 and moisture < 48): triggered.add(117)
    if (temp > 25 and light < 2500 and moisture < 48) != (temp > 25 and light < 8500 and moisture < 48): triggered.add(118)
    if (temp > 25 and light < 2500 and moisture < 48) != (temp > 25 and light < 2500 and moisture < 38): triggered.add(119)
    if (temp > 25 and light < 2500 and moisture < 48) != (temp > 25 and light < 2500 and moisture < 28): triggered.add(120)
    if (temp > 25 and light < 2500 and moisture < 48) != (temp < 25 and light < 2500 and moisture < 48): triggered.add(121)
    if (temp > 25 and light < 2500 and moisture < 48) != (temp > 25 and light > 2500 and moisture < 48): triggered.add(122)
    if (temp > 25 and light < 2500 and moisture < 48) != (temp > 25 and light < 2500 and moisture > 48): triggered.add(123)
    if (temp > 25 and light < 2500 and moisture < 48) != (temp > 25 or light < 2500 and moisture < 48): triggered.add(124)

    if (moisture < 52 and light > 2500 and light < 3500) != (moisture < 42 and light > 2500 and light < 3500): triggered.add(125)
    if (moisture < 52 and light > 2500 and light < 3500) != (moisture < 22 and light > 2500 and light < 3500): triggered.add(126)
    if (moisture < 52 and light > 2500 and light < 3500) != (moisture > 52 and light > 2500 and light < 3500): triggered.add(127)
    if (moisture < 52 and light > 2500 and light < 3500) != (moisture < 52 and light > 3000 and light < 3500): triggered.add(128)
    if (moisture < 52 and light > 2500 and light < 3500) != (moisture < 52 and light > 1500 and light < 3500): triggered.add(129)
    if (moisture < 52 and light > 2500 and light < 3500) != (moisture < 52 and light < 2500 and light < 3500): triggered.add(130)
    if (moisture < 52 and light > 2500 and light < 3500) != (moisture < 52 and light > 2500 and light > 3500): triggered.add(131)
    if (moisture < 52 and light > 2500 and light < 3500) != (moisture < 52 and light > 2500 and light < 5500): triggered.add(132)
    if (moisture < 52 and light > 2500 and light < 3500) != (moisture < 52 and light > 2500 and light < 6500): triggered.add(133)
    if (moisture < 52 and light > 2500 and light < 3500) != (moisture < 52 and light > 2500 and light < 7500): triggered.add(134)
    if (moisture < 52 and light > 2500 and light < 3500) != (moisture < 52 and light > 2500 and light < 3800): triggered.add(135)

    return triggered


def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """"""
    generated_path = execute_validation_rules(particle[0], particle[1], particle[2])

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # : x:1-50, y:1-50, z:1-50
        self.bounds = bounds if bounds else [(1, 100), (1, 100), (1, 100)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """target pathsPSO"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(particles[i][0], particles[i][1], particles[i][2]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(gbest_particle[0], gbest_particle[1], gbest_particle[2]),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """Path PSO"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - Path ")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations")
    print(f"Path : {len(target_paths)}")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"Path {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "" if result['success'] else f"({result['best_fitness']:.3f})"
        print(f"{status} | {result['time']:.2f}s | iterations{result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f": {success_count}/{len(target_paths)} ({success_rate:.1f}%) | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs}")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations, {len(target_paths)}Path ")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"---  {run_idx}/{num_runs} ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f": {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs} runcompleted | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 1: 
    ws1 = wb.active
    ws1.title = ""
    ws1.sheet_view.showGridLines = False

    headers = ["Run", "", "", "", "Average Iterations", "(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f" {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # 2: Path 
    ws2 = wb.create_sheet(title="Path ")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Path ID", "", "", "", "Average Iterations", "Minimum Iterations", "Maximum Iterations"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"Path  {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # 3: 
    ws3 = wb.create_sheet(title="")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Path ", "", "(x,y,z)", "", "Iterations", "Path "]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"{run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # 4: target paths
    ws4 = wb.create_sheet(title="target paths")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Path ID", "target paths", ""]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"Path  {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f" : {filename}")
    print(f"{'=' * 70}")
    print(f":")
    print(f"  1.        - {len(all_results)} run")
    print(f"  2. Path        - Path ")
    print(f"  3.    -  runPath ")
    print(f"  4. target paths       - target paths")
    print(f"{'=' * 70}\n")

    return filename


def main():


    target_paths = [
        {4, 14, 15, 16, 21, 22, 29, 32, 33, 38, 39, 40, 41, 43, 44, 48, 55, 56, 58, 59, 61, 62, 63, 64, 65, 69, 70, 71,
         73, 74, 75, 78, 79, 80, 82, 84, 89, 91, 93, 94, 95, 102, 103, 104, 112, 116, 117, 118, 122, 124, 125, 126, 127,
         128, 130, 131},
        {3, 4, 21, 22, 33, 38, 40, 41, 48, 55, 56, 58, 59, 61, 62, 63, 64, 65, 69, 70, 71, 73, 74, 78, 79, 80, 81, 82,
         85, 86, 87, 88, 90, 91, 92, 94, 95, 100, 102, 103, 104, 108, 109, 111, 112, 114, 119, 120, 121, 122, 123, 129,
         130},
        {4, 14, 15, 16, 18, 21, 22, 33, 38, 40, 41, 44, 48, 55, 56, 58, 59, 61, 62, 63, 64, 65, 69, 70, 71, 73, 74, 75,
         78, 79, 80, 81, 82, 84, 89, 91, 93, 94, 95, 102, 103, 104, 108, 109, 111, 112, 119, 120, 121, 122, 123, 129,
         130},
        {4, 14, 15, 16, 21, 22, 33, 38, 39, 40, 41, 44, 48, 55, 56, 58, 59, 61, 62, 63, 64, 65, 69, 70, 71, 73, 74, 75,
         78, 79, 80, 81, 82, 84, 91, 93, 94, 95, 102, 103, 104, 108, 109, 111, 112, 119, 120, 121, 122, 123, 129, 130},
        {4, 10, 14, 15, 16, 21, 33, 38, 41, 48, 55, 56, 58, 59, 61, 62, 63, 64, 65, 69, 70, 71, 73, 74, 75, 78, 79, 80,
         81, 82, 85, 86, 87, 88, 91, 92, 94, 95, 102, 103, 104, 108, 109, 111, 112, 119, 120, 121, 122, 123, 129, 130},
        {3, 4, 21, 22, 29, 32, 33, 38, 39, 40, 41, 43, 44, 55, 56, 58, 59, 61, 62, 63, 66, 67, 68, 70, 78, 79, 80, 82,
         84, 89, 91, 93, 94, 95, 100, 102, 103, 104, 112, 116, 117, 118, 122, 124, 125, 126, 127, 128, 130, 131},
        {3, 4, 9, 10, 14, 19, 21, 33, 38, 41, 54, 57, 58, 60, 64, 65, 66, 69, 73, 74, 75, 78, 79, 80, 81, 82, 85, 86,
         87, 88, 91, 92, 94, 95, 100, 102, 103, 104, 108, 109, 111, 112, 114, 119, 120, 121, 122, 123, 129, 130},
        {3, 4, 21, 22, 33, 38, 40, 41, 55, 56, 58, 59, 61, 62, 63, 66, 67, 68, 70, 78, 79, 80, 81, 82, 85, 86, 87, 88,
         90, 91, 92, 94, 95, 100, 102, 103, 104, 108, 109, 111, 112, 114, 119, 120, 121, 122, 123, 129, 130},
        {1, 2, 6, 7, 12, 13, 16, 17, 33, 35, 36, 37, 38, 66, 70, 82, 83, 91, 112, 131, 132, 133, 134},
        {3, 4, 6, 7, 12, 14, 19, 21, 22, 29, 32, 33, 34, 35, 36, 37, 38, 42, 55, 56, 58, 59, 61, 62, 63, 66, 67, 68, 70,
         75, 78, 79, 80, 82, 84, 89, 91, 93, 101, 104, 112, 125, 126, 127, 128, 130, 131},
        {3, 4, 10, 11, 12, 21, 22, 33, 38, 41, 55, 56, 58, 59, 61, 62, 63, 66, 67, 68, 70, 75, 78, 79, 80, 81, 82, 85,
         86, 87, 88, 91, 92, 101, 104, 108, 109, 111, 112, 115, 121, 124, 129, 130},
        {5, 21, 22, 23, 24, 25, 33, 38, 41, 55, 56, 58, 59, 62, 63, 76, 77, 79, 87, 88, 91, 92, 96, 97, 102, 113, 123,
         124},
        {5, 25, 26, 29, 30, 31, 34, 35, 36, 37, 38, 56, 58, 59, 62, 63, 85, 86, 87, 88, 90, 91, 92, 105, 106, 107, 111,
         112, 127},
        {2, 14, 15, 16, 18, 33, 37, 38, 45, 46, 48, 59, 60, 64, 65, 69, 70, 71, 73, 74, 82, 91, 99, 103, 112, 117, 118,
         122, 124, 131, 133, 134},
        {1, 2, 26, 27, 28, 33, 35, 36, 37, 38, 46, 48, 59, 60, 64, 72, 91, 107, 111, 112, 124, 131, 132, 133, 134},
        {14, 15, 16, 18, 33, 37, 38, 48, 49, 52, 59, 60, 64, 65, 69, 70, 71, 73, 74, 82, 91, 103, 112, 122, 124, 131},
        {1, 2, 14, 17, 19, 33, 35, 36, 37, 38, 58, 59, 61, 62, 63, 66, 68, 70, 82, 83, 91, 112, 131, 132, 133, 134},
        {1, 2, 6, 7, 12, 13, 16, 17, 20, 21, 22, 33, 35, 36, 37, 38, 58, 59, 61, 62, 63, 66, 68, 70, 82, 83, 91, 112,
         131, 132, 133, 134, 135},
        {1, 2, 33, 34, 35, 36, 37, 38, 46, 48, 58, 59, 61, 62, 63, 64, 65, 69, 70, 71, 73, 74, 82, 83, 91, 98, 99, 103,
         112, 116, 117, 118, 122, 124, 131, 132, 133, 134, 135},
        {4, 14, 15, 16, 18, 21, 22, 29, 32, 33, 34, 35, 36, 37, 38, 55, 56, 58, 59, 61, 62, 63, 66, 67, 68, 70, 75, 78,
         79, 80, 82, 84, 91, 93, 98, 99, 103, 112, 116, 117, 118, 122, 124, 126, 127, 128, 130, 131},
        {3, 4, 6, 7, 12, 14, 17, 19, 21, 22, 29, 32, 33, 38, 39, 40, 41, 43, 44, 55, 56, 58, 59, 61, 62, 63, 66, 67, 68,
         70, 78, 79, 80, 82, 84, 89, 91, 93, 104, 112, 125, 126, 127, 128, 130, 131},
        {3, 4, 20, 21, 22, 29, 32, 33, 34, 35, 36, 37, 38, 48, 56, 58, 59, 61, 62, 63, 64, 65, 69, 70, 71, 73, 74, 82,
         83, 84, 91, 93, 98, 99, 103, 112, 116, 117, 118, 122, 124, 125, 126, 127, 130, 131},
        {4, 10, 14, 15, 16, 21, 33, 38, 41, 48, 54, 57, 58, 60, 64, 65, 73, 74, 75, 78, 79, 80, 81, 82, 85, 86, 87, 88,
         91, 92, 94, 95, 102, 103, 104, 108, 109, 111, 112, 119, 120, 121, 122, 123, 129, 130},
        {4, 9, 10, 14, 15, 16, 18, 21, 33, 38, 41, 54, 57, 58, 60, 64, 65, 66, 69, 73, 74, 75, 78, 79, 80, 81, 82, 85,
         86, 87, 88, 91, 92, 101, 104, 108, 109, 111, 112, 115, 121, 124, 129, 130},
        {3, 4, 8, 14, 17, 19, 21, 33, 38, 41, 54, 57, 58, 60, 64, 65, 66, 69, 73, 74, 78, 79, 80, 81, 82, 87, 88, 91,
         92, 108, 109, 111, 112, 129, 130},
        {5, 20, 21, 22, 38, 41, 48, 55, 56, 58, 59, 62, 63, 64, 72, 77, 79, 88, 91, 97, 102, 110, 111, 112, 123, 124},
        {33, 37, 38, 47, 49, 50, 51, 53, 59, 60, 66, 70, 82, 91, 103, 112, 122, 124, 131}
    ]

    print("=" * 70)
    print("baseline PSO")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("Program completed")


if __name__ == "__main__":
    main()