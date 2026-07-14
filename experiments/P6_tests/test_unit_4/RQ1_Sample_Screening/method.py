import os
import random
from collections import deque
import numpy as np
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"


STATE_MIN_X, STATE_MAX_X = 1000, 10000
STATE_MIN_Y, STATE_MAX_Y = 38, 85
STATE_MIN_Z, STATE_MAX_Z = 1, 40
def generate_input():
    return [
        random.randint(STATE_MIN_X, STATE_MAX_X),
        random.randint(STATE_MIN_Y, STATE_MAX_Y),
        random.randint(STATE_MIN_Z, STATE_MAX_Z)
    ]



def execute_Tr(light, moisture, temp):
    actions = []
    triggered = set()

    if (light < 3500 and moisture < 55) != (light < 5500 and moisture < 55): triggered.add(1)
    if (light < 3500 and moisture < 55) != (light < 6500 and moisture < 55): triggered.add(2)
    if (light < 3500 and moisture < 55) != (light < 3500 and moisture < 45): triggered.add(3)
    if (light < 3500 and moisture < 55) != (light < 3500 and moisture < 35): triggered.add(4)
    if (light < 3500 and moisture < 55) != (light < 3500 and moisture < 75): triggered.add(5)

    if (light < 2200 and moisture < 48 and temp < 20) != (light < 4200 and moisture < 48 and temp < 20): triggered.add(6)
    if (light < 2200 and moisture < 48 and temp < 20) != (light < 6200 and moisture < 48 and temp < 20): triggered.add(7)
    if (light < 2200 and moisture < 48 and temp < 20) != (light < 2200 and moisture < 58 and temp < 20): triggered.add(8)
    if (light < 2200 and moisture < 48 and temp < 20) != (light < 2200 and moisture < 48 and temp < 30): triggered.add(9)
    if (light < 2200 and moisture < 48 and temp < 20) != (light < 2200 and moisture < 48 and temp > 20): triggered.add(10)
    if (light < 2200 and moisture < 48 and temp < 20) != (light < 2200 and moisture > 48 and temp < 20): triggered.add(11)
    if (light < 2200 and moisture < 48 and temp < 20) != (light > 2200 and moisture < 48 and temp < 20): triggered.add(12)

    if ((light < 1800 or temp < 18) and moisture < 45) != ((light < 1800 and temp < 18) and moisture < 45): triggered.add(13)
    if ((light < 1800 or temp < 18) and moisture < 45) != ((light < 1800 or temp < 18) or moisture < 45): triggered.add(14)
    if ((light < 1800 or temp < 18) and moisture < 45) != ((light > 1800 or temp < 18) and moisture < 45): triggered.add(15)
    if ((light < 1800 or temp < 18) and moisture < 45) != ((light < 1800 or temp > 18) and moisture < 45): triggered.add(16)
    if ((light < 1800 or temp < 18) and moisture < 45) != ((light < 1800 or temp < 18) and moisture > 45): triggered.add(17)
    if ((light < 1800 or temp < 18) and moisture < 45) != ((light < 1800 or temp < 38) and moisture < 45): triggered.add(18)
    if ((light < 1800 or temp < 18) and moisture < 45) != ((light < 1800 or temp < 18) and moisture < 65): triggered.add(19)

    if (light * moisture < 160000) != (light * moisture < 130000): triggered.add(20)
    if (light * moisture < 160000) != (light * moisture < 60000): triggered.add(21)
    if (light * moisture < 160000) != (light * moisture < 90000): triggered.add(22)

    if (light > 2500 and light < 3500 and moisture > 50 and moisture < 58) != (light > 1500 and light < 3500 and moisture > 50 and moisture < 58): triggered.add(23)
    if (light > 2500 and light < 3500 and moisture > 50 and moisture < 58) != (light > 500 and light < 3500 and moisture > 50 and moisture < 58): triggered.add(24)
    if (light > 2500 and light < 3500 and moisture > 50 and moisture < 58) != (light < 2500 and light < 3500 and moisture > 50 and moisture < 58): triggered.add(25)
    if (light > 2500 and light < 3500 and moisture > 50 and moisture < 58) != (light > 2500 and light > 3500 and moisture > 50 and moisture < 58): triggered.add(26)
    if (light > 2500 and light < 3500 and moisture > 50 and moisture < 58) != (light > 2500 and light < 6500 and moisture > 50 and moisture < 58): triggered.add(27)
    if (light > 2500 and light < 3500 and moisture > 50 and moisture < 58) != (light > 2500 and light < 8500 and moisture > 50 and moisture < 58): triggered.add(28)
    if (light > 2500 and light < 3500 and moisture > 50 and moisture < 58) != (light > 2500 and light < 3500 and moisture < 50 and moisture < 58): triggered.add(29)
    if (light > 2500 and light < 3500 and moisture > 50 and moisture < 58) != (light > 2500 and light < 3500 and moisture > 50 and moisture > 58): triggered.add(30)
    if (light > 2500 and light < 3500 and moisture > 50 and moisture < 58) != (light > 2500 and light < 3500 and moisture > 50 and moisture < 38): triggered.add(31)
    if (light > 2500 and light < 3500 and moisture > 50 and moisture < 58) != (light > 2500 and light < 3500 and moisture > 20 and moisture < 58): triggered.add(32)
    if (light > 2500 and light < 3500 and moisture > 50 and moisture < 58) != (light > 2500 and light < 3500 and moisture > 50 or moisture < 58): triggered.add(33)

    if ((3000 - light) > (58 - moisture) * 25) != ((4000 - light) > (58 - moisture) * 25): triggered.add(34)
    if ((3000 - light) > (58 - moisture) * 25) != ((5000 - light) > (58 - moisture) * 25): triggered.add(35)
    if ((3000 - light) > (58 - moisture) * 25) != ((6000 - light) > (58 - moisture) * 25): triggered.add(36)
    if ((3000 - light) > (58 - moisture) * 25) != ((3000 + light) > (58 - moisture) * 25): triggered.add(37)
    if ((3000 - light) > (58 - moisture) * 25) != ((3000 - light) < (58 - moisture) * 25): triggered.add(38)
    if ((3000 - light) > (58 - moisture) * 25) != ((3000 - light) > (68 - moisture) * 25): triggered.add(39)
    if ((3000 - light) > (58 - moisture) * 25) != ((3000 - light) > (78 - moisture) * 25): triggered.add(40)
    if ((3000 - light) > (58 - moisture) * 25) != ((3000 - light) > (58 + moisture) * 25): triggered.add(41)
    if ((3000 - light) > (58 - moisture) * 25) != ((3000 - light) > (58 - moisture) * 15): triggered.add(42)
    if ((3000 - light) > (58 - moisture) * 25) != ((3000 - light) > (58 - moisture) * 35): triggered.add(43)
    if ((3000 - light) > (58 - moisture) * 25) != ((3000 - light) > (58 - moisture) * 45): triggered.add(44)

    if (light > 8500 and temp > 30) != (light > 5500 and temp > 30): triggered.add(45)
    if (light > 8500 and temp > 30) != (light > 3500 and temp > 30): triggered.add(46)
    if (light > 8500 and temp > 30) != (light > 8500 and temp > 20): triggered.add(47)
    if (light > 8500 and temp > 30) != (light < 8500 and temp > 30): triggered.add(48)
    if (light > 8500 and temp > 30) != (light > 8500 and temp < 30): triggered.add(49)
    if (light > 8500 and temp > 30) != (light > 8500 and moisture > 30): triggered.add(50)
    if (light > 8500 and temp > 30) != (light > 8500 and temp + moisture > 30): triggered.add(51)
    if (light > 8500 and temp > 30) != (light > 8500 and moisture - temp > 30): triggered.add(52)
    if (light > 8500 and temp > 30) != (light > 8500 and temp > 15): triggered.add(53)

    if (light > 2000 and light < 4000 and moisture > 40 and moisture < 65) != (light > 1000 and light < 4000 and moisture > 40 and moisture < 65): triggered.add(54)
    if (light > 2000 and light < 4000 and moisture > 40 and moisture < 65) != (light > 3000 and light < 4000 and moisture > 40 and moisture < 65): triggered.add(55)
    if (light > 2000 and light < 4000 and moisture > 40 and moisture < 65) != (light > 3500 and light < 4000 and moisture > 40 and moisture < 65): triggered.add(56)
    if (light > 2000 and light < 4000 and moisture > 40 and moisture < 65) != (light > 1500 and light < 4000 and moisture > 40 and moisture < 65): triggered.add(57)
    if (light > 2000 and light < 4000 and moisture > 40 and moisture < 65) != (light < 2000 and light < 4000 and moisture > 40 and moisture < 65): triggered.add(58)
    if (light > 2000 and light < 4000 and moisture > 40 and moisture < 65) != (light > 2000 and light > 4000 and moisture > 40 and moisture < 65): triggered.add(59)
    if (light > 2000 and light < 4000 and moisture > 40 and moisture < 65) != (light > 2000 and light < 4000 or moisture > 40 and moisture < 65): triggered.add(60)
    if (light > 2000 and light < 4000 and moisture > 40 and moisture < 65) != (light > 2000 and light < 4000 and moisture > 50 and moisture < 65): triggered.add(61)
    if (light > 2000 and light < 4000 and moisture > 40 and moisture < 65) != (light > 2000 and light < 4000 and moisture > 40 and moisture > 65): triggered.add(62)
    if (light > 2000 and light < 4000 and moisture > 40 and moisture < 65) != (light > 2000 and light < 4000 and moisture < 40 and moisture < 65): triggered.add(63)

    if (moisture < 50 and (light < 2000 or temp > 30)) != (moisture > 50 and (light < 2000 or temp > 30)): triggered.add(64)
    if (moisture < 50 and (light < 2000 or temp > 30)) != (moisture < 20 and (light < 2000 or temp > 30)): triggered.add(65)
    if (moisture < 50 and (light < 2000 or temp > 30)) != (moisture < 50 and (light > 2000 or temp > 30)): triggered.add(66)
    if (moisture < 50 and (light < 2000 or temp > 30)) != (moisture < 50 and (light < 3000 or temp > 30)): triggered.add(67)
    if (moisture < 50 and (light < 2000 or temp > 30)) != (moisture < 50 and (light < 4000 or temp > 30)): triggered.add(68)
    if (moisture < 50 and (light < 2000 or temp > 30)) != (moisture < 50 and (light < 2000 and temp > 30)): triggered.add(69)
    if (moisture < 50 and (light < 2000 or temp > 30)) != (moisture < 50 and (light < 2000 or temp < 30)): triggered.add(70)
    if (moisture < 50 and (light < 2000 or temp > 30)) != (moisture < 50 and (light < 2000 or temp > 50)): triggered.add(71)
    if (moisture < 50 and (light < 2000 or temp > 30)) != (moisture < 70 and (light < 2000 or temp > 30)): triggered.add(72)
    if (moisture < 50 and (light < 2000 or temp > 30)) != (moisture < 40 and (light < 2000 or temp > 30)): triggered.add(73)
    if (moisture < 50 and (light < 2000 or temp > 30)) != (moisture < 25 and (light < 2000 or temp > 30)): triggered.add(74)

    if (moisture > 35 and moisture < 50 and light > 1500 and light < 3000) != (moisture > 45 and moisture < 50 and light > 1500 and light < 3000): triggered.add(75)
    if (moisture > 35 and moisture < 50 and light > 1500 and light < 3000) != (moisture > 35 and moisture < 60 and light > 1500 and light < 3000): triggered.add(76)
    if (moisture > 35 and moisture < 50 and light > 1500 and light < 3000) != (moisture > 35 and moisture < 70 and light > 1500 and light < 3000): triggered.add(77)
    if (moisture > 35 and moisture < 50 and light > 1500 and light < 3000) != (moisture < 35 and moisture < 50 and light > 1500 and light < 3000): triggered.add(78)
    if (moisture > 35 and moisture < 50 and light > 1500 and light < 3000) != (moisture > 35 and moisture > 50 and light > 1500 and light < 3000): triggered.add(79)
    if (moisture > 35 and moisture < 50 and light > 1500 and light < 3000) != (moisture > 35 and moisture < 50 and light < 1500 and light < 3000): triggered.add(80)
    if (moisture > 35 and moisture < 50 and light > 1500 and light < 3000) != (moisture > 35 and moisture < 50 and light > 2500 and light < 3000): triggered.add(81)
    if (moisture > 35 and moisture < 50 and light > 1500 and light < 3000) != (moisture > 35 and moisture < 50 and light > 1500 and light > 3000): triggered.add(82)
    if (moisture > 35 and moisture < 50 and light > 1500 and light < 3000) != (moisture > 35 and moisture < 50 and light > 1500 and light < 5000): triggered.add(83)

    if (moisture < 40 + (light - 2000) / 100) != (moisture < 40 + (light - 3000) / 100): triggered.add(84)
    if (moisture < 40 + (light - 2000) / 100) != (moisture < 40 + (light - 1000) / 100): triggered.add(85)
    if (moisture < 40 + (light - 2000) / 100) != (moisture < 50 + (light - 2000) / 100): triggered.add(86)
    if (moisture < 40 + (light - 2000) / 100) != (moisture < 60 + (light - 2000) / 100): triggered.add(87)
    if (moisture < 40 + (light - 2000) / 100) != (moisture < 40 + (light + 2000) / 100): triggered.add(88)
    if (moisture < 40 + (light - 2000) / 100) != (moisture < 40 + (light - 2000) / 200): triggered.add(89)
    if (moisture < 40 + (light - 2000) / 100) != (moisture < 40 + (light - 2000) / 50): triggered.add(90)
    if (moisture < 40 + (light - 2000) / 100) != (moisture > 40 + (light - 2000) / 100): triggered.add(91)
    if (moisture < 40 + (light - 2000) / 100) != (moisture < 60 + (light - 2000) / 100): triggered.add(92)
    if (moisture < 40 + (light - 2000) / 100) != (moisture < 30 + (light - 2000) / 100): triggered.add(93)

    if (moisture < 48 and light < 2800 and temp > 24) != (moisture < 38 and light < 2800 and temp > 24): triggered.add(94)
    if (moisture < 48 and light < 2800 and temp > 24) != (moisture < 28 and light < 2800 and temp > 24): triggered.add(95)
    if (moisture < 48 and light < 2800 and temp > 24) != (moisture < 58 and light < 2800 and temp > 24): triggered.add(96)
    if (moisture < 48 and light < 2800 and temp > 24) != (moisture < 68 and light < 2800 and temp > 24): triggered.add(97)
    if (moisture < 48 and light < 2800 and temp > 24) != (moisture < 48 and light < 3800 and temp > 24): triggered.add(98)
    if (moisture < 48 and light < 2800 and temp > 24) != (moisture < 48 and light < 5800 and temp > 24): triggered.add(99)
    if (moisture < 48 and light < 2800 and temp > 24) != (moisture < 48 and light < 2800 and temp > 34): triggered.add(100)
    if (moisture < 48 and light < 2800 and temp > 24) != (moisture < 48 and light < 2800 and temp > 14): triggered.add(101)
    if (moisture < 48 and light < 2800 and temp > 24) != (moisture > 48 and light < 2800 and temp > 24): triggered.add(102)
    if (moisture < 48 and light < 2800 and temp > 24) != (moisture < 48 and light > 2800 and temp > 24): triggered.add(103)
    if (moisture < 48 and light < 2800 and temp > 24) != (moisture < 48 and light < 2800 and temp < 24): triggered.add(104)

    if (light < 2500 and moisture > 50 and moisture < 60) != (light < 3500 and moisture > 50 and moisture < 60): triggered.add(105)
    if (light < 2500 and moisture > 50 and moisture < 60) != (light < 4500 and moisture > 50 and moisture < 60): triggered.add(106)
    if (light < 2500 and moisture > 50 and moisture < 60) != (light < 6500 and moisture > 50 and moisture < 60): triggered.add(107)
    if (light < 2500 and moisture > 50 and moisture < 60) != (light < 2500 and moisture > 30 and moisture < 60): triggered.add(108)
    if (light < 2500 and moisture > 50 and moisture < 60) != (light < 2500 and moisture > 20 and moisture < 60): triggered.add(109)
    if (light < 2500 and moisture > 50 and moisture < 60) != (light < 2500 and moisture > 50 and moisture < 70): triggered.add(110)
    if (light < 2500 and moisture > 50 and moisture < 60) != (light < 2500 or moisture > 50 and moisture < 60): triggered.add(111)
    if (light < 2500 and moisture > 50 and moisture < 60) != (light < 2500 and moisture > 50 or moisture < 60): triggered.add(112)
    if (light < 2500 and moisture > 50 and moisture < 60) != (light < 2000 and moisture > 50 and moisture < 60): triggered.add(113)

    if (temp > 25 and light < 2500 and moisture < 48) != (temp > 35 and light < 2500 and moisture < 48): triggered.add(114)
    if (temp > 25 and light < 2500 and moisture < 48) != (temp > 15 and light < 2500 and moisture < 48): triggered.add(115)
    if (temp > 25 and light < 2500 and moisture < 48) != (temp > 25 and light < 4500 and moisture < 48): triggered.add(116)
    if (temp > 25 and light < 2500 and moisture < 48) != (temp > 25 and light < 6500 and moisture < 48): triggered.add(117)
    if (temp > 25 and light < 2500 and moisture < 48) != (temp > 25 and light < 8500 and moisture < 48): triggered.add(118)
    if (temp > 25 and light < 2500 and moisture < 48) != (temp > 25 and light < 2500 and moisture < 38): triggered.add(119)
    if (temp > 25 and light < 2500 and moisture < 48) != (temp > 25 and light < 2500 and moisture < 28): triggered.add(120)
    if (temp > 25 and light < 2500 and moisture < 48) != (temp < 25 and light < 2500 and moisture < 48): triggered.add(121)
    if (temp > 25 and light < 2500 and moisture < 48) != (temp > 25 and light > 2500 and moisture < 48): triggered.add(122)
    if (temp > 25 and light < 2500 and moisture < 48) != (temp > 25 and light < 2500 and moisture > 48): triggered.add(123)
    if (temp > 25 and light < 2500 and moisture < 48) != (temp > 25 or light < 2500 and moisture < 48): triggered.add(124)

    if (moisture < 52 and light > 2500 and light < 3500) != (moisture < 42 and light > 2500 and light < 3500): triggered.add(125)
    if (moisture < 52 and light > 2500 and light < 3500) != (moisture < 22 and light > 2500 and light < 3500): triggered.add(126)
    if (moisture < 52 and light > 2500 and light < 3500) != (moisture > 52 and light > 2500 and light < 3500): triggered.add(127)
    if (moisture < 52 and light > 2500 and light < 3500) != (moisture < 52 and light > 3000 and light < 3500): triggered.add(128)
    if (moisture < 52 and light > 2500 and light < 3500) != (moisture < 52 and light > 1500 and light < 3500): triggered.add(129)
    if (moisture < 52 and light > 2500 and light < 3500) != (moisture < 52 and light < 2500 and light < 3500): triggered.add(130)
    if (moisture < 52 and light > 2500 and light < 3500) != (moisture < 52 and light > 2500 and light > 3500): triggered.add(131)
    if (moisture < 52 and light > 2500 and light < 3500) != (moisture < 52 and light > 2500 and light < 5500): triggered.add(132)
    if (moisture < 52 and light > 2500 and light < 3500) != (moisture < 52 and light > 2500 and light < 6500): triggered.add(133)
    if (moisture < 52 and light > 2500 and light < 3500) != (moisture < 52 and light > 2500 and light < 7500): triggered.add(134)
    if (moisture < 52 and light > 2500 and light < 3500) != (moisture < 52 and light > 2500 and light < 3800): triggered.add(135)

    return triggered

def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0


targetPaths = [
    {4, 14, 15, 16, 21, 22, 29, 32, 33, 38, 39, 40, 41, 43, 44, 48, 55, 56, 58, 59, 61, 62, 63, 64, 65, 69, 70, 71,
     73, 74, 75, 78, 79, 80, 82, 84, 89, 91, 93, 94, 95, 102, 103, 104, 112, 116, 117, 118, 122, 124, 125, 126, 127,
     128, 130, 131},
    {3, 4, 21, 22, 33, 38, 40, 41, 48, 55, 56, 58, 59, 61, 62, 63, 64, 65, 69, 70, 71, 73, 74, 78, 79, 80, 81, 82,
     85, 86, 87, 88, 90, 91, 92, 94, 95, 100, 102, 103, 104, 108, 109, 111, 112, 114, 119, 120, 121, 122, 123, 129,
     130},
    {4, 14, 15, 16, 18, 21, 22, 33, 38, 40, 41, 44, 48, 55, 56, 58, 59, 61, 62, 63, 64, 65, 69, 70, 71, 73, 74, 75,
     78, 79, 80, 81, 82, 84, 89, 91, 93, 94, 95, 102, 103, 104, 108, 109, 111, 112, 119, 120, 121, 122, 123, 129,
     130},
    {4, 14, 15, 16, 21, 22, 33, 38, 39, 40, 41, 44, 48, 55, 56, 58, 59, 61, 62, 63, 64, 65, 69, 70, 71, 73, 74, 75,
     78, 79, 80, 81, 82, 84, 91, 93, 94, 95, 102, 103, 104, 108, 109, 111, 112, 119, 120, 121, 122, 123, 129, 130},
    {4, 10, 14, 15, 16, 21, 33, 38, 41, 48, 55, 56, 58, 59, 61, 62, 63, 64, 65, 69, 70, 71, 73, 74, 75, 78, 79, 80,
     81, 82, 85, 86, 87, 88, 91, 92, 94, 95, 102, 103, 104, 108, 109, 111, 112, 119, 120, 121, 122, 123, 129, 130},
    {3, 4, 21, 22, 29, 32, 33, 38, 39, 40, 41, 43, 44, 55, 56, 58, 59, 61, 62, 63, 66, 67, 68, 70, 78, 79, 80, 82,
     84, 89, 91, 93, 94, 95, 100, 102, 103, 104, 112, 116, 117, 118, 122, 124, 125, 126, 127, 128, 130, 131},
    {3, 4, 9, 10, 14, 19, 21, 33, 38, 41, 54, 57, 58, 60, 64, 65, 66, 69, 73, 74, 75, 78, 79, 80, 81, 82, 85, 86,
     87, 88, 91, 92, 94, 95, 100, 102, 103, 104, 108, 109, 111, 112, 114, 119, 120, 121, 122, 123, 129, 130},
    {3, 4, 21, 22, 33, 38, 40, 41, 55, 56, 58, 59, 61, 62, 63, 66, 67, 68, 70, 78, 79, 80, 81, 82, 85, 86, 87, 88,
     90, 91, 92, 94, 95, 100, 102, 103, 104, 108, 109, 111, 112, 114, 119, 120, 121, 122, 123, 129, 130},
    {1, 2, 6, 7, 12, 13, 16, 17, 33, 35, 36, 37, 38, 66, 70, 82, 83, 91, 112, 131, 132, 133, 134},
    {3, 4, 6, 7, 12, 14, 19, 21, 22, 29, 32, 33, 34, 35, 36, 37, 38, 42, 55, 56, 58, 59, 61, 62, 63, 66, 67, 68, 70,
     75, 78, 79, 80, 82, 84, 89, 91, 93, 101, 104, 112, 125, 126, 127, 128, 130, 131},
    {3, 4, 10, 11, 12, 21, 22, 33, 38, 41, 55, 56, 58, 59, 61, 62, 63, 66, 67, 68, 70, 75, 78, 79, 80, 81, 82, 85,
     86, 87, 88, 91, 92, 101, 104, 108, 109, 111, 112, 115, 121, 124, 129, 130},
    {5, 21, 22, 23, 24, 25, 33, 38, 41, 55, 56, 58, 59, 62, 63, 76, 77, 79, 87, 88, 91, 92, 96, 97, 102, 113, 123,
     124},
    {5, 25, 26, 29, 30, 31, 34, 35, 36, 37, 38, 56, 58, 59, 62, 63, 85, 86, 87, 88, 90, 91, 92, 105, 106, 107, 111,
     112, 127},
    {2, 14, 15, 16, 18, 33, 37, 38, 45, 46, 48, 59, 60, 64, 65, 69, 70, 71, 73, 74, 82, 91, 99, 103, 112, 117, 118,
     122, 124, 131, 133, 134},
    {1, 2, 26, 27, 28, 33, 35, 36, 37, 38, 46, 48, 59, 60, 64, 72, 91, 107, 111, 112, 124, 131, 132, 133, 134},
    {14, 15, 16, 18, 33, 37, 38, 48, 49, 52, 59, 60, 64, 65, 69, 70, 71, 73, 74, 82, 91, 103, 112, 122, 124, 131},
    {1, 2, 14, 17, 19, 33, 35, 36, 37, 38, 58, 59, 61, 62, 63, 66, 68, 70, 82, 83, 91, 112, 131, 132, 133, 134},
    {1, 2, 6, 7, 12, 13, 16, 17, 20, 21, 22, 33, 35, 36, 37, 38, 58, 59, 61, 62, 63, 66, 68, 70, 82, 83, 91, 112,
     131, 132, 133, 134, 135},
    {1, 2, 33, 34, 35, 36, 37, 38, 46, 48, 58, 59, 61, 62, 63, 64, 65, 69, 70, 71, 73, 74, 82, 83, 91, 98, 99, 103,
     112, 116, 117, 118, 122, 124, 131, 132, 133, 134, 135},
    {4, 14, 15, 16, 18, 21, 22, 29, 32, 33, 34, 35, 36, 37, 38, 55, 56, 58, 59, 61, 62, 63, 66, 67, 68, 70, 75, 78,
     79, 80, 82, 84, 91, 93, 98, 99, 103, 112, 116, 117, 118, 122, 124, 126, 127, 128, 130, 131},
    {3, 4, 6, 7, 12, 14, 17, 19, 21, 22, 29, 32, 33, 38, 39, 40, 41, 43, 44, 55, 56, 58, 59, 61, 62, 63, 66, 67, 68,
     70, 78, 79, 80, 82, 84, 89, 91, 93, 104, 112, 125, 126, 127, 128, 130, 131},
    {3, 4, 20, 21, 22, 29, 32, 33, 34, 35, 36, 37, 38, 48, 56, 58, 59, 61, 62, 63, 64, 65, 69, 70, 71, 73, 74, 82,
     83, 84, 91, 93, 98, 99, 103, 112, 116, 117, 118, 122, 124, 125, 126, 127, 130, 131},
    {4, 10, 14, 15, 16, 21, 33, 38, 41, 48, 54, 57, 58, 60, 64, 65, 73, 74, 75, 78, 79, 80, 81, 82, 85, 86, 87, 88,
     91, 92, 94, 95, 102, 103, 104, 108, 109, 111, 112, 119, 120, 121, 122, 123, 129, 130},
    {4, 9, 10, 14, 15, 16, 18, 21, 33, 38, 41, 54, 57, 58, 60, 64, 65, 66, 69, 73, 74, 75, 78, 79, 80, 81, 82, 85,
     86, 87, 88, 91, 92, 101, 104, 108, 109, 111, 112, 115, 121, 124, 129, 130},
    {3, 4, 8, 14, 17, 19, 21, 33, 38, 41, 54, 57, 58, 60, 64, 65, 66, 69, 73, 74, 78, 79, 80, 81, 82, 87, 88, 91,
     92, 108, 109, 111, 112, 129, 130},
    {5, 20, 21, 22, 38, 41, 48, 55, 56, 58, 59, 62, 63, 64, 72, 77, 79, 88, 91, 97, 102, 110, 111, 112, 123, 124},
    {33, 37, 38, 47, 49, 50, 51, 53, 59, 60, 66, 70, 82, 91, 103, 112, 122, 124, 131}
]


class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 2000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(targetPaths)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }


def compute_robustness(state, path):
    base = execute_Tr(*state)
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                        np.clip(state[0] + dx, STATE_MIN_X, STATE_MAX_X),
                        np.clip(state[1] + dy, STATE_MIN_Y, STATE_MAX_Y),
                        np.clip(state[2] + dz, STATE_MIN_Z, STATE_MAX_Z)
                    ])
                n_trig = execute_Tr(*neighbor)
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = targetPaths[target_path_idx]
    samples = []
    attempts = 0

    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(STATE_MIN_X, STATE_MAX_X),
            random.randint(STATE_MIN_Y, STATE_MAX_Y),
            random.randint(STATE_MIN_Z, STATE_MAX_Z)
        ])
        triggered = execute_Tr(*state)

        if not triggered:
            continue

        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)

        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }

        samples.append(sample_data)

    return samples


def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample[
                'robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)

        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]

    return selected_samples


def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = targetPaths[target_path_idx]
        samples = []
        attempts = 0

        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(STATE_MIN_X, STATE_MAX_X),
                random.randint(STATE_MIN_Y, STATE_MAX_Y),
                random.randint(STATE_MIN_Z, STATE_MAX_Z)
            ])
            triggered = execute_Tr(*state)

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)

            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)

        random.shuffle(samples)
        return samples[:config.top_k_samples]

    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")

        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)


def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}

    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates

    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}

        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config,
                                                         shared_candidates[path_idx])

            strategy_results[path_idx] = samples

        results[strategy_name] = strategy_results

    return results


def analyze_fitness_values(results, config):
    analysis_results = {}

    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []

        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])

        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }

        all_scores = []

        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])

        scores_array = np.array(all_scores)

        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)

        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })

        analysis_results[strategy_name] = analysis

    return analysis_results


def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []

    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)

    df = pd.DataFrame(df_data)
    return df, analysis_results


def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []

    print(f"Starting {num_runs} experiments...")

    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")

        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)

        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)

        print(f"Experiment {run_idx + 1} completed.")

    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df


def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')

    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width

    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []

    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]

        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)

    stats_df = pd.DataFrame(stats_data)

    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Results saved to: {output_path}")


def main():
    results_df = run_multiple_experiments(num_runs=20)

    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")

    save_results_to_excel(results_df, output_path)

    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)

    return results_df, output_path


if __name__ == "__main__":
    results, output_file = main()
