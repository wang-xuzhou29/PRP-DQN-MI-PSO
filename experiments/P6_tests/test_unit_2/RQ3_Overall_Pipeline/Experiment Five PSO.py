import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

def execute_validation_rules(x, y, z):
    triggered = set()

    # 固定参数
    energy_price = 0.15
    time_of_day = 12
    temp = 25
    humidity = 60
    energy_trend = "stable"
    weather_forecast = "sunny"

    # 分支1-10: 能量价格相关规则
    if (energy_price > 0.10 and time_of_day == 12 and (x + temp) > 500) != (energy_price > 0.10 and time_of_day == 12 and (x + temp) > 800):
        triggered.add(1)
    if (energy_price > 0.10 and time_of_day == 12 and (x + temp) > 1000) != (energy_price > 0.10 and time_of_day == 12 and (x + temp) > 900):
        triggered.add(2)
    if (energy_price < 0.20 and (x + z) < 800) != (energy_price < 0.20 and (x + z) < 1200):
        triggered.add(3)
    if (energy_price < 0.20 and (x + z) < 600) != (energy_price < 0.20 and (x + z) < 1000):
        triggered.add(4)
    if (energy_price > 0.10 and (x * temp) > 20000) != (energy_price > 0.10 and (x * temp) > 30000):
        triggered.add(5)
    if (energy_price > 0.10 and (x * temp) > 15000) != (energy_price > 0.10 and (x * temp) > 25000):
        triggered.add(6)
    if (energy_price < 0.20 and (y + humidity) < 100) != (energy_price < 0.20 and (y + humidity) < 150):
        triggered.add(7)
    if (energy_price < 0.20 and (y + humidity) < 80) != (energy_price < 0.20 and (y + humidity) < 120):
        triggered.add(8)
    if (energy_price > 0.10 and z > 300) != (energy_price > 0.10 and z > 450):
        triggered.add(9)
    if (energy_price > 0.10 and z > 200) != (energy_price > 0.10 and z > 350):
        triggered.add(10)

    # 分支11-19: 时间段相关规则
    if (time_of_day == 12 and x < 500 and y > 200) != (time_of_day == 12 and x < 700 and y > 200):
        triggered.add(11)
    if (time_of_day == 12 and x < 400 and z < 300) != (time_of_day == 12 and x < 600 and z < 300):
        triggered.add(12)
    if (temp + humidity > 20 and y < 500) != (temp + humidity < 20 and y < 500):
        triggered.add(13)
    if (energy_price < 0.20 and (z + x) < 800 and y > 100) != (energy_price < 0.20 and (z + x) < 1200 and y > 100):
        triggered.add(14)
    if (energy_price < 0.20 and (z + x) < 600 and y < 800) != (energy_price < 0.20 and (z + x) < 1000 and y < 800):
        triggered.add(15)
    if (temp < 30 and x > 500 and z > 200) != (temp < 30 and x < 500 and z > 200):
        triggered.add(16)
    if (temp < 28 and x > 400 and y > 300) != (temp < 28 and x > 400 and z > 300):
        triggered.add(17)
    if (time_of_day == 12 and x > 600 and z < 400) != (time_of_day == 12 and x > 800 and z < 400):
        triggered.add(18)
    if (time_of_day == 12 and x > 500 and y + z > 600) != (time_of_day == 12 and x > 750 and y + z > 600):
        triggered.add(19)

    # 分支20-29: 更多组合规则
    if (time_of_day == 12 and z < 300 and x > 200) != (time_of_day == 12 and z < 450 and x > 200):
        triggered.add(20)
    if (time_of_day == 12 and z < 250 and y > 300) != (time_of_day == 12 and z < 400 and y > 300):
        triggered.add(21)
    if (energy_price > 0.10 and x < 400 and y + z > 500) != (energy_price > 0.10 and x < 600 and y + z > 500):
        triggered.add(22)
    if (energy_price > 0.10 and x < 300 and z > 200) != (energy_price > 0.10 and x < 500 and z > 200):
        triggered.add(23)
    if (time_of_day == 12 and temp > 20 and x > 400) != (time_of_day == 12 and temp > 30 and x > 400):
        triggered.add(24)
    if (time_of_day == 12 and temp > 18 and y < 600) != (time_of_day == 12 and temp > 28 and y < 600):
        triggered.add(25)
    if (humidity > 50 and y > 500 and x < 800) != (humidity > 70 and y > 500 and x < 800):
        triggered.add(26)
    if (humidity > 45 and y > 450 and z < 400) != (humidity > 65 and y > 450 and z < 400):
        triggered.add(27)
    if (time_of_day == 12 and y < 200 and x + z > 500) != (time_of_day == 12 and y < 300 and x + z > 500):
        triggered.add(28)
    if (time_of_day == 12 and y < 150 and z > 100) != (time_of_day == 12 and y < 250 and z > 100):
        triggered.add(29)

    # 分支30-37: 更多规则
    if (energy_price > 0.10 and x > 700) != (energy_price > 0.10 and x > 850):
        triggered.add(30)
    if (energy_price > 0.10 and x > 600) != (energy_price > 0.10 and x > 750):
        triggered.add(31)
    if (temp < 28 and y < 250) != (temp < 28 and y < 450):
        triggered.add(32)
    if (energy_price > 0.10 and z > 350) != (energy_price > 0.10 and x > 350):
        triggered.add(33)
    if (humidity < 70 and y < 400) != (humidity < 30 and y < 400):
        triggered.add(34)
    if (humidity < 65 and y < 350) != (humidity < 65 and x < 350):
        triggered.add(35)
    if (time_of_day == 12 and (x + temp) < 600) != (time_of_day == 12 and (x + temp) < 800):
        triggered.add(36)
    if (time_of_day == 12 and (x + temp) < 500) != (time_of_day == 12 and (x + temp) < 700):
        triggered.add(37)

    # 分支38-53: 复杂表达式
    if (energy_price > 0.10 and (x * 0.15) > 80 and temp > 20) != (energy_price > 0.10 and (x * 0.15) > 150 and temp > 20):
        triggered.add(38)
    if (energy_price > 0.10 and (x * 0.15) > 60 and temp > 20) != (energy_price > 0.10 and (x * 0.15) > 120 and temp > 20):
        triggered.add(39)
    if (energy_price < 0.20 and (z * 0.5) > 100 and x < 600) != (energy_price < 0.20 and (z * 0.5) > 200 and x < 600):
        triggered.add(40)
    if (energy_price < 0.20 and (z * 0.5) > 80 and x < 600) != (energy_price < 0.20 and (z * 0.5) > 180 and x < 600):
        triggered.add(41)
    if (temp + humidity > 80 and y > 400) != (temp + humidity > 100 and y > 400):
        triggered.add(42)
    if (temp + humidity > 70 and y > 350) != (temp + humidity > 90 and y > 350):
        triggered.add(43)
    if (energy_price < 0.20 and (y + z) < 400) != (energy_price < 0.20 and (y + z) < 600):
        triggered.add(44)
    if (energy_price < 0.20 and (y + z) < 350) != (energy_price < 0.20 and (y + z) < 550):
        triggered.add(45)
    if (time_of_day == 12 and (x * temp) > 15000) != (time_of_day == 12 and (x * temp) > 25000):
        triggered.add(46)
    if (time_of_day == 12 and (x * temp) > 12000) != (time_of_day == 12 and (x * temp) > 20000):
        triggered.add(47)
    if (time_of_day == 12 and (z + temp) < 400) != (time_of_day == 12 and (z + temp) < 600):
        triggered.add(48)
    if (time_of_day == 12 and (z + temp) < 350) != (time_of_day == 12 and (z + temp) < 550):
        triggered.add(49)
    if (energy_price > 0.10 and (humidity * 0.15) > 8 and y > 70) != (energy_price > 0.10 and (humidity * 0.15) > 8 and y > 870):
        triggered.add(50)
    if (energy_price < 0.20 and (x * y * 0.01) < 200 and z < 300) != (energy_price < 0.20 and (x * y * 0.01) < 400 and z < 300):
        triggered.add(51)
    if (time_of_day == 12 and (humidity + y) * 0.5 < 150) != (time_of_day == 12 and (humidity + y) * 0.5 < 250):
        triggered.add(52)
    if (time_of_day == 12 and (humidity + y) * 0.5 < 120) != (time_of_day == 12 and (humidity + y) * 0.5 < 220):
        triggered.add(53)

    # 分支54-63: energy_trend相关
    if (energy_trend == "stable" and x < 500 and y > 200) != (energy_trend == "stable" and x < 700 and y > 200):
        triggered.add(54)
    if (energy_trend == "stable" and x < 400 and z < 300) != (energy_trend == "stable" and x < 600 and z < 300):
        triggered.add(55)
    if (energy_price > 0.10 and temp > 20 and y > 300) != (energy_price > 0.10 and temp > 30 and y > 300):
        triggered.add(56)
    if (energy_price > 0.10 and temp > 18 and z < 400) != (energy_price > 0.10 and temp > 28 and z < 400):
        triggered.add(57)
    if (energy_trend == "stable" and z < 300 and x + y > 600) != (energy_trend == "stable" and z < 450 and x + y > 600):
        triggered.add(58)
    if (energy_trend == "stable" and z < 250 and y > 400) != (energy_trend == "stable" and z < 400 and y > 400):
        triggered.add(59)
    if (energy_trend == "stable" and y < 300 and x > 500) != (energy_trend == "stable" and y < 450 and x > 500):
        triggered.add(60)
    if (energy_trend == "stable" and y < 250 and z > 200) != (energy_trend == "stable" and y < 400 and z > 200):
        triggered.add(61)
    if (humidity > 50 and x > 600 and y + z > 700) != (humidity > 70 and x > 600 and y + z > 700):
        triggered.add(62)
    if (humidity > 45 and x > 550 and z < 400) != (humidity > 65 and x > 550 and z < 400):
        triggered.add(63)

    # 分支64-71: 更多组合
    if (energy_trend == "stable" and time_of_day == 12 and x > 600) != (energy_trend == "stable" and time_of_day == 12 and x > 800):
        triggered.add(64)
    if (energy_trend == "stable" and time_of_day == 12 and x > 500) != (energy_trend == "stable" and time_of_day == 12 and x > 750):
        triggered.add(65)
    if (energy_price < 0.20 and z > 350) != (energy_price < 0.20 and z > 450):
        triggered.add(66)
    if (energy_price < 0.20 and z > 300) != (energy_price < 0.20 and z > 400):
        triggered.add(67)
    if (energy_trend == "stable" and (x + temp) < 600) != (energy_trend == "stable" and (x + temp) < 800):
        triggered.add(68)
    if (energy_trend == "stable" and (x + temp) < 500) != (energy_trend == "stable" and (x + temp) < 700):
        triggered.add(69)
    if (humidity < 70 and z < 400) != (humidity < 70 and x < 400):
        triggered.add(70)
    if (humidity < 65 and z < 350) != (humidity < 65 and z < 750):
        triggered.add(71)

    # 分支72-81: weather_forecast相关
    if (weather_forecast == "sunny" and x < 600 and y > 300) != (weather_forecast == "sunny" and x < 800 and y > 300):
        triggered.add(72)
    if (weather_forecast == "sunny" and x < 500 and z < 400) != (weather_forecast == "sunny" and x < 700 and z < 400):
        triggered.add(73)
    if (weather_forecast == "sunny" and y < 500 and x > 200) != (weather_forecast == "sunny" and y < 700 and x > 200):
        triggered.add(74)
    if (weather_forecast == "sunny" and y < 400 and z > 250) != (weather_forecast == "sunny" and y < 600 and z > 250):
        triggered.add(75)
    if (weather_forecast == "sunny" and humidity > 50 and x + y > 800) != (weather_forecast == "sunny" and humidity > 70 and x + y > 800):
        triggered.add(76)
    if (weather_forecast == "sunny" and humidity > 45 and z < 300) != (weather_forecast == "sunny" and humidity > 65 and z < 300):
        triggered.add(77)
    if (weather_forecast == "sunny" and temp > 20 and x * y > 200000) != (weather_forecast == "sunny" and temp > 30 and x * y > 200000):
        triggered.add(78)
    if (weather_forecast == "sunny" and temp > 18 and y + z < 800) != (weather_forecast == "sunny" and temp > 28 and y + z < 800):
        triggered.add(79)
    if (weather_forecast == "sunny" and z < 300 and x > 400) != (weather_forecast == "sunny" and z < 450 and x > 400):
        triggered.add(80)
    if (weather_forecast == "sunny" and z < 250 and y < 600) != (weather_forecast == "sunny" and z < 400 and y < 600):
        triggered.add(81)

    # 分支82-85: 高级组合
    if (weather_forecast == "sunny" and time_of_day == 12 and y < 300 and x > 400) != (weather_forecast == "sunny" and time_of_day == 12 and y < 500 and x > 400):
        triggered.add(82)
    if (weather_forecast == "sunny" and time_of_day == 12 and y < 250 and z > 200) != (weather_forecast == "sunny" and time_of_day == 12 and y < 450 and z > 200):
        triggered.add(83)
    if (weather_forecast == "sunny" and energy_price > 0.10 and x + y > 1000) != (weather_forecast == "sunny" and energy_price > 0.20 and x + y > 1000):
        triggered.add(84)
    if (weather_forecast == "sunny" and energy_price > 0.05 and z < 400) != (weather_forecast == "sunny" and energy_price > 0.18 and z < 400):
        triggered.add(85)

    # 分支86-100: 最复杂的规则
    if (energy_price * 10 + time_of_day * 0.5 > 7 and x > 600 and z > 200) != (energy_price * 10 + time_of_day * 0.5 > 11 and x > 600 and z > 200):
        triggered.add(86)
    if (energy_price * temp > 3 and humidity > 50 and x * y > 300000) != (energy_price * temp > 5 and humidity > 50 and x * y > 300000):
        triggered.add(87)
    if (energy_price * temp > 2.5 and humidity > 50 and z < 400) != (energy_price * temp > 4.5 and humidity > 50 and z < 400):
        triggered.add(88)
    if ((1 - energy_price * 2) * (x + z) > 800 and y < 400) != ((1 - energy_price * 2) * (x + z) > 1200 and y < 400):
        triggered.add(89)
    if ((1 - energy_price * 2) * (x + z) > 600 and y < 400) != ((1 - energy_price * 2) * (x + z) > 1000 and y < 400):
        triggered.add(90)
    if (time_of_day == 12 and x * temp * 0.1 > 2000 and y + z > 600) != (time_of_day == 12 and x * temp * 0.1 > 3000 and y + z > 600):
        triggered.add(91)
    if (time_of_day == 12 and x * temp * 0.1 > 1500 and z < 500) != (time_of_day == 12 and x * temp * 0.1 > 2500 and z < 500):
        triggered.add(92)
    if (time_of_day == 12 and (z + y) * 0.8 < 400 and x > 300) != (time_of_day == 12 and (z + y) * 0.8 < 600 and x > 300):
        triggered.add(93)
    if (time_of_day == 12 and (z + y) * 0.8 < 350 and x < 700) != (time_of_day == 12 and (z + y) * 0.8 < 550 and x < 700):
        triggered.add(94)
    if (energy_price * 10 + humidity * 0.1 > 7 and temp > 20 and z > x / 2) != (energy_price * 10 + humidity * 0.1 > 11 and temp > 20 and z > x / 2):
        triggered.add(95)
    if ((1 - energy_price * 3) * (x + temp) > 500 and z < 300 and x + y > 800) != ((1 - energy_price * 3) * (x + temp) > 800 and z < 300 and x + y > 800):
        triggered.add(96)
    if (energy_price > 0.10 and abs(time_of_day - 12) < 1 and (x + temp) > 600 and y * z > 100000) != (energy_price > 0.10 and abs(time_of_day - 12) < 1 and (x + temp) > 900 and y * z > 100000):
        triggered.add(97)
    if (energy_price > 0.10 and abs(time_of_day - 12) < 1 and (x + temp) > 500 and y + z < 1000) != (energy_price > 0.10 and abs(time_of_day - 12) < 1 and (x + temp) > 800 and y + z < 1000):
        triggered.add(98)
    if (24 - time_of_day > 8 and (y + humidity) < 400 and x > z) != (24 - time_of_day > 8 and (y + humidity) < 600 and x > z):
        triggered.add(99)
    if (24 - time_of_day > 8 and (y + humidity) < 350 and x + y + z > 1000) != (24 - time_of_day > 8 and (y + humidity) < 550 and x + y + z > 1000):
        triggered.add(100)

    return triggered


def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """"""
    generated_path = execute_validation_rules(particle[0], particle[1], particle[2])

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # : x:1-50, y:1-50, z:1-50
        self.bounds = bounds if bounds else [(1, 100), (1, 100), (1, 100)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """target pathsPSO"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(particles[i][0], particles[i][1], particles[i][2]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(gbest_particle[0], gbest_particle[1], gbest_particle[2]),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """Path PSO"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - Path ")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations")
    print(f"Path : {len(target_paths)}")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"Path {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "" if result['success'] else f"({result['best_fitness']:.3f})"
        print(f"{status} | {result['time']:.2f}s | iterations{result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f": {success_count}/{len(target_paths)} ({success_rate:.1f}%) | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs}")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations, {len(target_paths)}Path ")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"---  {run_idx}/{num_runs} ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f": {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs} runcompleted | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 1: 
    ws1 = wb.active
    ws1.title = ""
    ws1.sheet_view.showGridLines = False

    headers = ["Run", "", "", "", "Average Iterations", "(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f" {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # 2: Path 
    ws2 = wb.create_sheet(title="Path ")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Path ID", "", "", "", "Average Iterations", "Minimum Iterations", "Maximum Iterations"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"Path  {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # 3: 
    ws3 = wb.create_sheet(title="")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Path ", "", "(x,y,z)", "", "Iterations", "Path "]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"{run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # 4: target paths
    ws4 = wb.create_sheet(title="target paths")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Path ID", "target paths", ""]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"Path  {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f" : {filename}")
    print(f"{'=' * 70}")
    print(f":")
    print(f"  1.        - {len(all_results)} run")
    print(f"  2. Path        - Path ")
    print(f"  3.    -  runPath ")
    print(f"  4. target paths       - target paths")
    print(f"{'=' * 70}\n")

    return filename


def main():


    target_paths = [
        [1, 3, 6, 9, 11, 13, 14, 16, 18, 19, 20, 21, 24, 25, 31, 32, 34, 36, 37, 38, 39, 43, 46, 47, 48, 49, 50,
         52, 53, 54, 56, 57, 58, 60, 61, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 76, 78, 79, 80, 81, 82,
         83, 85, 86, 88, 90, 92, 93, 95, 97, 98, 99, 100],
        [1, 3, 4, 6, 9, 11, 13, 14, 15, 16, 18, 19, 20, 21, 24, 25, 31, 32, 34, 36, 37, 38, 39, 43, 46, 47, 49, 50, 52,
         54, 56, 57, 58, 60, 61, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 76, 78, 79, 80, 81, 82, 83, 84, 85,
         86, 88, 90, 92, 95, 97, 98, 99, 100],
        [1, 3, 4, 6, 9, 10, 11, 13, 14, 15, 16, 18, 19, 20, 21, 24, 25, 31, 32, 33, 36, 37, 38, 39, 42, 43, 46, 47, 50,
         52, 54, 56, 57, 58, 59, 60, 62, 63, 64, 65, 67, 68, 69, 70, 72, 73, 75, 76, 78, 79, 80, 81, 82, 83, 84, 85,
         86, 88, 92, 93, 95, 97, 98, 99, 100],
        [1, 3, 4, 6, 9, 10, 11, 13, 14, 15, 16, 18, 19, 20, 21, 24, 25, 31, 32, 33, 34, 35, 36, 37, 38, 39, 46, 47, 49,
         50, 52, 53, 54, 56, 57, 58, 60, 61, 63, 64, 65, 67, 68, 69, 70, 72, 73, 76, 78, 79, 80, 81, 82, 83, 85, 86,
         88, 90, 92, 93, 94, 95, 97, 98, 100],
        [1, 3, 4, 6, 10, 11, 13, 14, 15, 16, 17, 18, 19, 20, 21, 24, 25, 31, 32, 33, 34, 36, 37, 38, 39, 43, 46, 47,
         50, 52, 53, 54, 56, 57, 58, 60, 61, 63, 64, 65, 68, 69, 70, 72, 73, 76, 78, 79, 80, 81, 82, 83, 84, 85, 86,
         88, 90, 92, 93, 94, 97, 98, 99, 100],
        [1, 3, 4, 6, 10, 11, 13, 14, 15, 16, 17, 18, 19, 21, 24, 25, 31, 32, 33, 34, 36, 37, 38, 39, 43, 46, 47, 50,
         52, 53, 54, 56, 57, 60, 61, 63, 64, 65, 68, 69, 70, 72, 73, 76, 77, 78, 79, 81, 82, 83, 84, 85, 86, 88, 90,
         92, 93, 94, 97, 98, 99, 100],
        [1, 3, 4, 6, 9, 10, 11, 13, 14, 15, 16, 18, 19, 20, 21, 24, 25, 27, 31, 33, 36, 37, 38, 39, 42, 43, 46, 47, 49,
         50, 54, 56, 57, 58, 59, 62, 63, 64, 65, 67, 68, 69, 70, 72, 73, 75, 76, 78, 80, 81, 82, 84, 85, 86, 87, 88,
         92, 95, 97, 98, 99, 100],
        [1, 3, 4, 6, 9, 11, 14, 15, 16, 18, 19, 20, 21, 24, 25, 26, 27, 31, 36, 37, 38, 39, 42, 43, 46, 47, 48, 49, 50,
         54, 56, 57, 58, 59, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 75, 76, 78, 80, 81, 84, 85, 86, 87,
         88, 92, 95, 97, 98, 99],
        [1, 3, 6, 9, 13, 14, 16, 18, 19, 20, 21, 24, 25, 30, 31, 32, 34, 35, 36, 38, 39, 46, 47, 48, 49, 50, 52, 53,
         56, 57, 58, 60, 61, 62, 63, 64, 65, 66, 67, 68, 70, 71, 72, 76, 78, 79, 80, 81, 82, 83, 84, 85, 86, 88, 90,
         92, 93, 95, 97, 98, 100],
        [1, 3, 4, 6, 9, 11, 13, 14, 15, 16, 17, 18, 19, 20, 24, 25, 28, 31, 32, 34, 35, 36, 37, 38, 39, 46, 47, 49,
         50, 52, 53, 54, 57, 58, 61, 63, 64, 65, 66, 67, 68, 69, 70, 71, 73, 76, 79, 80, 81, 83, 85, 86, 88, 90, 92,
         93, 94, 95, 97, 98, 100],
        [1, 3, 4, 9, 10, 11, 13, 14, 15, 16, 19, 20, 21, 22, 24, 25, 32, 33, 34, 36, 37, 38, 39, 40, 41, 43, 47, 49,
         50, 52, 53, 54, 56, 57, 58, 60, 61, 63, 65, 67, 68, 69, 70, 73, 76, 78, 79, 80, 81, 82, 83, 85, 88, 90, 93,
         95, 97, 98, 99, 100],
        [1, 3, 4, 6, 10, 11, 13, 14, 15, 16, 17, 18, 24, 25, 31, 32, 33, 34, 35, 36, 37, 38, 39, 44, 45, 46, 47, 50,
         52, 53, 54, 56, 57, 60, 61, 63, 64, 65, 68, 69, 70, 72, 73, 76, 77, 78, 79, 82, 83, 85, 86, 88, 90, 92, 93,
         94, 98, 100],
        [1, 3, 4, 10, 11, 12, 13, 14, 15, 16, 17, 19, 21, 22, 24, 25, 32, 33, 34, 36, 37, 38, 39, 40, 41, 43, 47, 50,
         52, 54, 55, 56, 57, 60, 61, 63, 65, 68, 69, 70, 73, 76, 77, 78, 79, 81, 82, 83, 85, 88, 90, 93, 94, 97, 98,
         99, 100],
        [1, 3, 4, 6, 9, 10, 11, 13, 14, 15, 16, 17, 18, 20, 24, 25, 28, 29, 31, 33, 34, 35, 36, 37, 38, 39, 44, 46,
         47, 50, 52, 53, 54, 57, 58, 63, 64, 65, 67, 68, 69, 70, 73, 76, 79, 80, 81, 85, 86, 88, 90, 92, 93, 94, 98],
        [3, 6, 9, 13, 14, 16, 18, 20, 21, 24, 25, 30, 32, 34, 38, 39, 43, 46, 47, 48, 49, 50, 52, 53, 56, 57, 58, 60,
         61, 62, 63, 64, 66, 67, 70, 71, 72, 76, 78, 79, 80, 81, 82, 83, 84, 85, 86, 88, 89, 90, 92, 97, 99, 100],
        [2, 5, 6, 9, 13, 16, 20, 21, 24, 25, 32, 34, 38, 43, 46, 48, 49, 50, 52, 53, 56, 57, 58, 60, 61, 62, 63, 66,
         67, 70, 71, 76, 78, 79, 80, 81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 99, 100],
        [1, 3, 4, 9, 13, 14, 15, 16, 20, 21, 22, 23, 24, 25, 32, 34, 37, 39, 40, 43, 47, 48, 49, 50, 52, 53, 56, 57,
         58, 61, 66, 67, 69, 70, 71, 76, 79, 80, 81, 82, 83, 85, 88, 90, 93, 95, 98, 99, 100],
        [1, 3, 6, 7, 9, 13, 16, 17, 18, 20, 24, 25, 31, 34, 35, 36, 37, 38, 39, 44, 45, 46, 47, 48, 49, 50, 57, 58,
         63, 64, 65, 66, 67, 68, 69, 70, 71, 73, 79, 80, 81, 85, 86, 88, 90, 92, 94, 95, 98],
        [1, 3, 6, 7, 8, 9, 13, 16, 17, 18, 20, 24, 25, 31, 34, 35, 36, 37, 38, 39, 44, 45, 46, 47, 48, 49, 57, 58, 63,
         64, 65, 66, 67, 68, 69, 70, 71, 73, 79, 80, 81, 85, 86, 88, 90, 92, 95, 98],
        [2, 3, 5, 6, 10, 13, 14, 16, 17, 21, 24, 25, 32, 33, 34, 38, 43, 46, 50, 52, 53, 56, 57, 60, 61, 63, 70, 76,
         77, 78, 79, 81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 96, 99, 100],
        [1, 3, 4, 6, 8, 10, 13, 15, 16, 18, 24, 25, 31, 33, 34, 35, 36, 37, 38, 39, 46, 47, 51, 57, 63, 64, 65, 68,
         69, 70, 73, 77, 79, 85, 86, 88, 90, 92, 98]
    ]

    print("=" * 70)
    print("baseline PSO")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("Program completed")


if __name__ == "__main__":
    main()