import os
import random
import numpy as np
import time
import psutil
import math
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

NUM_RUNS = 20  # 默认实验运行次数

# === 新的三维范围设置 ===
LIGHT_MIN = 1
LIGHT_MAX = 1000
MOISTURE_MIN = 1
MOISTURE_MAX = 1000
TEMP_MIN = 1
TEMP_MAX = 600

BOUNDS = {
    'light': (LIGHT_MIN, LIGHT_MAX),
    'temp': (TEMP_MIN, TEMP_MAX),
    'moisture': (MOISTURE_MIN, MOISTURE_MAX)
}


def clip_state(state):
    """将状态限制在各维度的边界内"""
    return np.array([
        np.clip(state[0], BOUNDS['light'][0], BOUNDS['light'][1]),
        np.clip(state[1], BOUNDS['temp'][0], BOUNDS['temp'][1]),
        np.clip(state[2], BOUNDS['moisture'][0], BOUNDS['moisture'][1])
    ])


def execute_Tr(position):
    """执行目标函数并返回触发的路径"""
    x = int(np.clip(position[0], BOUNDS['light'][0], BOUNDS['light'][1]))
    temp = int(np.clip(position[1], BOUNDS['temp'][0], BOUNDS['temp'][1]))
    z = int(np.clip(position[2], BOUNDS['moisture'][0], BOUNDS['moisture'][1]))
    return category1_multivariable_control(x, temp, z)


# === 目标函数 ===
def category1_multivariable_control(x, y, z):
    triggered = set()

    # 固定参数
    energy_price = 0.15
    time_of_day = 12
    temp = 25
    humidity = 60
    energy_trend = "stable"
    weather_forecast = "sunny"

    # 分支1-10: 能量价格相关规则
    if (energy_price > 0.10 and time_of_day == 12 and (x + temp) > 500) != (energy_price > 0.10 and time_of_day == 12 and (x + temp) > 800):
        triggered.add(1)
    if (energy_price > 0.10 and time_of_day == 12 and (x + temp) > 1000) != (energy_price > 0.10 and time_of_day == 12 and (x + temp) > 900):
        triggered.add(2)
    if (energy_price < 0.20 and (x + z) < 800) != (energy_price < 0.20 and (x + z) < 1200):
        triggered.add(3)
    if (energy_price < 0.20 and (x + z) < 600) != (energy_price < 0.20 and (x + z) < 1000):
        triggered.add(4)
    if (energy_price > 0.10 and (x * temp) > 20000) != (energy_price > 0.10 and (x * temp) > 30000):
        triggered.add(5)
    if (energy_price > 0.10 and (x * temp) > 15000) != (energy_price > 0.10 and (x * temp) > 25000):
        triggered.add(6)
    if (energy_price < 0.20 and (y + humidity) < 100) != (energy_price < 0.20 and (y + humidity) < 150):
        triggered.add(7)
    if (energy_price < 0.20 and (y + humidity) < 80) != (energy_price < 0.20 and (y + humidity) < 120):
        triggered.add(8)
    if (energy_price > 0.10 and z > 300) != (energy_price > 0.10 and z > 450):
        triggered.add(9)
    if (energy_price > 0.10 and z > 200) != (energy_price > 0.10 and z > 350):
        triggered.add(10)

    # 分支11-19: 时间段相关规则
    if (time_of_day == 12 and x < 500 and y > 200) != (time_of_day == 12 and x < 700 and y > 200):
        triggered.add(11)
    if (time_of_day == 12 and x < 400 and z < 300) != (time_of_day == 12 and x < 600 and z < 300):
        triggered.add(12)
    if (temp + humidity > 20 and y < 500) != (temp + humidity < 20 and y < 500):
        triggered.add(13)
    if (energy_price < 0.20 and (z + x) < 800 and y > 100) != (energy_price < 0.20 and (z + x) < 1200 and y > 100):
        triggered.add(14)
    if (energy_price < 0.20 and (z + x) < 600 and y < 800) != (energy_price < 0.20 and (z + x) < 1000 and y < 800):
        triggered.add(15)
    if (temp < 30 and x > 500 and z > 200) != (temp < 30 and x < 500 and z > 200):
        triggered.add(16)
    if (temp < 28 and x > 400 and y > 300) != (temp < 28 and x > 400 and z > 300):
        triggered.add(17)
    if (time_of_day == 12 and x > 600 and z < 400) != (time_of_day == 12 and x > 800 and z < 400):
        triggered.add(18)
    if (time_of_day == 12 and x > 500 and y + z > 600) != (time_of_day == 12 and x > 750 and y + z > 600):
        triggered.add(19)

    # 分支20-29: 更多组合规则
    if (time_of_day == 12 and z < 300 and x > 200) != (time_of_day == 12 and z < 450 and x > 200):
        triggered.add(20)
    if (time_of_day == 12 and z < 250 and y > 300) != (time_of_day == 12 and z < 400 and y > 300):
        triggered.add(21)
    if (energy_price > 0.10 and x < 400 and y + z > 500) != (energy_price > 0.10 and x < 600 and y + z > 500):
        triggered.add(22)
    if (energy_price > 0.10 and x < 300 and z > 200) != (energy_price > 0.10 and x < 500 and z > 200):
        triggered.add(23)
    if (time_of_day == 12 and temp > 20 and x > 400) != (time_of_day == 12 and temp > 30 and x > 400):
        triggered.add(24)
    if (time_of_day == 12 and temp > 18 and y < 600) != (time_of_day == 12 and temp > 28 and y < 600):
        triggered.add(25)
    if (humidity > 50 and y > 500 and x < 800) != (humidity > 70 and y > 500 and x < 800):
        triggered.add(26)
    if (humidity > 45 and y > 450 and z < 400) != (humidity > 65 and y > 450 and z < 400):
        triggered.add(27)
    if (time_of_day == 12 and y < 200 and x + z > 500) != (time_of_day == 12 and y < 300 and x + z > 500):
        triggered.add(28)
    if (time_of_day == 12 and y < 150 and z > 100) != (time_of_day == 12 and y < 250 and z > 100):
        triggered.add(29)

    # 分支30-37: 更多规则
    if (energy_price > 0.10 and x > 700) != (energy_price > 0.10 and x > 850):
        triggered.add(30)
    if (energy_price > 0.10 and x > 600) != (energy_price > 0.10 and x > 750):
        triggered.add(31)
    if (temp < 28 and y < 250) != (temp < 28 and y < 450):
        triggered.add(32)
    if (energy_price > 0.10 and z > 350) != (energy_price > 0.10 and x > 350):
        triggered.add(33)
    if (humidity < 70 and y < 400) != (humidity < 30 and y < 400):
        triggered.add(34)
    if (humidity < 65 and y < 350) != (humidity < 65 and x < 350):
        triggered.add(35)
    if (time_of_day == 12 and (x + temp) < 600) != (time_of_day == 12 and (x + temp) < 800):
        triggered.add(36)
    if (time_of_day == 12 and (x + temp) < 500) != (time_of_day == 12 and (x + temp) < 700):
        triggered.add(37)

    # 分支38-53: 复杂表达式
    if (energy_price > 0.10 and (x * 0.15) > 80 and temp > 20) != (energy_price > 0.10 and (x * 0.15) > 150 and temp > 20):
        triggered.add(38)
    if (energy_price > 0.10 and (x * 0.15) > 60 and temp > 20) != (energy_price > 0.10 and (x * 0.15) > 120 and temp > 20):
        triggered.add(39)
    if (energy_price < 0.20 and (z * 0.5) > 100 and x < 600) != (energy_price < 0.20 and (z * 0.5) > 200 and x < 600):
        triggered.add(40)
    if (energy_price < 0.20 and (z * 0.5) > 80 and x < 600) != (energy_price < 0.20 and (z * 0.5) > 180 and x < 600):
        triggered.add(41)
    if (temp + humidity > 80 and y > 400) != (temp + humidity > 100 and y > 400):
        triggered.add(42)
    if (temp + humidity > 70 and y > 350) != (temp + humidity > 90 and y > 350):
        triggered.add(43)
    if (energy_price < 0.20 and (y + z) < 400) != (energy_price < 0.20 and (y + z) < 600):
        triggered.add(44)
    if (energy_price < 0.20 and (y + z) < 350) != (energy_price < 0.20 and (y + z) < 550):
        triggered.add(45)
    if (time_of_day == 12 and (x * temp) > 15000) != (time_of_day == 12 and (x * temp) > 25000):
        triggered.add(46)
    if (time_of_day == 12 and (x * temp) > 12000) != (time_of_day == 12 and (x * temp) > 20000):
        triggered.add(47)
    if (time_of_day == 12 and (z + temp) < 400) != (time_of_day == 12 and (z + temp) < 600):
        triggered.add(48)
    if (time_of_day == 12 and (z + temp) < 350) != (time_of_day == 12 and (z + temp) < 550):
        triggered.add(49)
    if (energy_price > 0.10 and (humidity * 0.15) > 8 and y > 70) != (energy_price > 0.10 and (humidity * 0.15) > 8 and y > 870):
        triggered.add(50)
    if (energy_price < 0.20 and (x * y * 0.01) < 200 and z < 300) != (energy_price < 0.20 and (x * y * 0.01) < 400 and z < 300):
        triggered.add(51)
    if (time_of_day == 12 and (humidity + y) * 0.5 < 150) != (time_of_day == 12 and (humidity + y) * 0.5 < 250):
        triggered.add(52)
    if (time_of_day == 12 and (humidity + y) * 0.5 < 120) != (time_of_day == 12 and (humidity + y) * 0.5 < 220):
        triggered.add(53)

    # 分支54-63: energy_trend相关
    if (energy_trend == "stable" and x < 500 and y > 200) != (energy_trend == "stable" and x < 700 and y > 200):
        triggered.add(54)
    if (energy_trend == "stable" and x < 400 and z < 300) != (energy_trend == "stable" and x < 600 and z < 300):
        triggered.add(55)
    if (energy_price > 0.10 and temp > 20 and y > 300) != (energy_price > 0.10 and temp > 30 and y > 300):
        triggered.add(56)
    if (energy_price > 0.10 and temp > 18 and z < 400) != (energy_price > 0.10 and temp > 28 and z < 400):
        triggered.add(57)
    if (energy_trend == "stable" and z < 300 and x + y > 600) != (energy_trend == "stable" and z < 450 and x + y > 600):
        triggered.add(58)
    if (energy_trend == "stable" and z < 250 and y > 400) != (energy_trend == "stable" and z < 400 and y > 400):
        triggered.add(59)
    if (energy_trend == "stable" and y < 300 and x > 500) != (energy_trend == "stable" and y < 450 and x > 500):
        triggered.add(60)
    if (energy_trend == "stable" and y < 250 and z > 200) != (energy_trend == "stable" and y < 400 and z > 200):
        triggered.add(61)
    if (humidity > 50 and x > 600 and y + z > 700) != (humidity > 70 and x > 600 and y + z > 700):
        triggered.add(62)
    if (humidity > 45 and x > 550 and z < 400) != (humidity > 65 and x > 550 and z < 400):
        triggered.add(63)

    # 分支64-71: 更多组合
    if (energy_trend == "stable" and time_of_day == 12 and x > 600) != (energy_trend == "stable" and time_of_day == 12 and x > 800):
        triggered.add(64)
    if (energy_trend == "stable" and time_of_day == 12 and x > 500) != (energy_trend == "stable" and time_of_day == 12 and x > 750):
        triggered.add(65)
    if (energy_price < 0.20 and z > 350) != (energy_price < 0.20 and z > 450):
        triggered.add(66)
    if (energy_price < 0.20 and z > 300) != (energy_price < 0.20 and z > 400):
        triggered.add(67)
    if (energy_trend == "stable" and (x + temp) < 600) != (energy_trend == "stable" and (x + temp) < 800):
        triggered.add(68)
    if (energy_trend == "stable" and (x + temp) < 500) != (energy_trend == "stable" and (x + temp) < 700):
        triggered.add(69)
    if (humidity < 70 and z < 400) != (humidity < 70 and x < 400):
        triggered.add(70)
    if (humidity < 65 and z < 350) != (humidity < 65 and z < 750):
        triggered.add(71)

    # 分支72-81: weather_forecast相关
    if (weather_forecast == "sunny" and x < 600 and y > 300) != (weather_forecast == "sunny" and x < 800 and y > 300):
        triggered.add(72)
    if (weather_forecast == "sunny" and x < 500 and z < 400) != (weather_forecast == "sunny" and x < 700 and z < 400):
        triggered.add(73)
    if (weather_forecast == "sunny" and y < 500 and x > 200) != (weather_forecast == "sunny" and y < 700 and x > 200):
        triggered.add(74)
    if (weather_forecast == "sunny" and y < 400 and z > 250) != (weather_forecast == "sunny" and y < 600 and z > 250):
        triggered.add(75)
    if (weather_forecast == "sunny" and humidity > 50 and x + y > 800) != (weather_forecast == "sunny" and humidity > 70 and x + y > 800):
        triggered.add(76)
    if (weather_forecast == "sunny" and humidity > 45 and z < 300) != (weather_forecast == "sunny" and humidity > 65 and z < 300):
        triggered.add(77)
    if (weather_forecast == "sunny" and temp > 20 and x * y > 200000) != (weather_forecast == "sunny" and temp > 30 and x * y > 200000):
        triggered.add(78)
    if (weather_forecast == "sunny" and temp > 18 and y + z < 800) != (weather_forecast == "sunny" and temp > 28 and y + z < 800):
        triggered.add(79)
    if (weather_forecast == "sunny" and z < 300 and x > 400) != (weather_forecast == "sunny" and z < 450 and x > 400):
        triggered.add(80)
    if (weather_forecast == "sunny" and z < 250 and y < 600) != (weather_forecast == "sunny" and z < 400 and y < 600):
        triggered.add(81)

    # 分支82-85: 高级组合
    if (weather_forecast == "sunny" and time_of_day == 12 and y < 300 and x > 400) != (weather_forecast == "sunny" and time_of_day == 12 and y < 500 and x > 400):
        triggered.add(82)
    if (weather_forecast == "sunny" and time_of_day == 12 and y < 250 and z > 200) != (weather_forecast == "sunny" and time_of_day == 12 and y < 450 and z > 200):
        triggered.add(83)
    if (weather_forecast == "sunny" and energy_price > 0.10 and x + y > 1000) != (weather_forecast == "sunny" and energy_price > 0.20 and x + y > 1000):
        triggered.add(84)
    if (weather_forecast == "sunny" and energy_price > 0.05 and z < 400) != (weather_forecast == "sunny" and energy_price > 0.18 and z < 400):
        triggered.add(85)

    # 分支86-100: 最复杂的规则
    if (energy_price * 10 + time_of_day * 0.5 > 7 and x > 600 and z > 200) != (energy_price * 10 + time_of_day * 0.5 > 11 and x > 600 and z > 200):
        triggered.add(86)
    if (energy_price * temp > 3 and humidity > 50 and x * y > 300000) != (energy_price * temp > 5 and humidity > 50 and x * y > 300000):
        triggered.add(87)
    if (energy_price * temp > 2.5 and humidity > 50 and z < 400) != (energy_price * temp > 4.5 and humidity > 50 and z < 400):
        triggered.add(88)
    if ((1 - energy_price * 2) * (x + z) > 800 and y < 400) != ((1 - energy_price * 2) * (x + z) > 1200 and y < 400):
        triggered.add(89)
    if ((1 - energy_price * 2) * (x + z) > 600 and y < 400) != ((1 - energy_price * 2) * (x + z) > 1000 and y < 400):
        triggered.add(90)
    if (time_of_day == 12 and x * temp * 0.1 > 2000 and y + z > 600) != (time_of_day == 12 and x * temp * 0.1 > 3000 and y + z > 600):
        triggered.add(91)
    if (time_of_day == 12 and x * temp * 0.1 > 1500 and z < 500) != (time_of_day == 12 and x * temp * 0.1 > 2500 and z < 500):
        triggered.add(92)
    if (time_of_day == 12 and (z + y) * 0.8 < 400 and x > 300) != (time_of_day == 12 and (z + y) * 0.8 < 600 and x > 300):
        triggered.add(93)
    if (time_of_day == 12 and (z + y) * 0.8 < 350 and x < 700) != (time_of_day == 12 and (z + y) * 0.8 < 550 and x < 700):
        triggered.add(94)
    if (energy_price * 10 + humidity * 0.1 > 7 and temp > 20 and z > x / 2) != (energy_price * 10 + humidity * 0.1 > 11 and temp > 20 and z > x / 2):
        triggered.add(95)
    if ((1 - energy_price * 3) * (x + temp) > 500 and z < 300 and x + y > 800) != ((1 - energy_price * 3) * (x + temp) > 800 and z < 300 and x + y > 800):
        triggered.add(96)
    if (energy_price > 0.10 and abs(time_of_day - 12) < 1 and (x + temp) > 600 and y * z > 100000) != (energy_price > 0.10 and abs(time_of_day - 12) < 1 and (x + temp) > 900 and y * z > 100000):
        triggered.add(97)
    if (energy_price > 0.10 and abs(time_of_day - 12) < 1 and (x + temp) > 500 and y + z < 1000) != (energy_price > 0.10 and abs(time_of_day - 12) < 1 and (x + temp) > 800 and y + z < 1000):
        triggered.add(98)
    if (24 - time_of_day > 8 and (y + humidity) < 400 and x > z) != (24 - time_of_day > 8 and (y + humidity) < 600 and x > z):
        triggered.add(99)
    if (24 - time_of_day > 8 and (y + humidity) < 350 and x + y + z > 1000) != (24 - time_of_day > 8 and (y + humidity) < 550 and x + y + z > 1000):
        triggered.add(100)

    return triggered


# 目标路径定义
targetPaths = [
    [1, 3, 6, 9, 11, 13, 14, 16, 18, 19, 20, 21, 24, 25, 31, 32, 34, 36, 37, 38, 39, 43, 46, 47, 48, 49, 50,
     52, 53, 54, 56, 57, 58, 60, 61, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 76, 78, 79, 80, 81, 82,
     83, 85, 86, 88, 90, 92, 93, 95, 97, 98, 99, 100],
    [1, 3, 4, 6, 9, 11, 13, 14, 15, 16, 18, 19, 20, 21, 24, 25, 31, 32, 34, 36, 37, 38, 39, 43, 46, 47, 49, 50, 52,
     54, 56, 57, 58, 60, 61, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 76, 78, 79, 80, 81, 82, 83, 84, 85,
     86, 88, 90, 92, 95, 97, 98, 99, 100],
    [1, 3, 4, 6, 9, 10, 11, 13, 14, 15, 16, 18, 19, 20, 21, 24, 25, 31, 32, 33, 36, 37, 38, 39, 42, 43, 46, 47, 50,
     52, 54, 56, 57, 58, 59, 60, 62, 63, 64, 65, 67, 68, 69, 70, 72, 73, 75, 76, 78, 79, 80, 81, 82, 83, 84, 85,
     86, 88, 92, 93, 95, 97, 98, 99, 100],
    [1, 3, 4, 6, 9, 10, 11, 13, 14, 15, 16, 18, 19, 20, 21, 24, 25, 31, 32, 33, 34, 35, 36, 37, 38, 39, 46, 47, 49,
     50, 52, 53, 54, 56, 57, 58, 60, 61, 63, 64, 65, 67, 68, 69, 70, 72, 73, 76, 78, 79, 80, 81, 82, 83, 85, 86,
     88, 90, 92, 93, 94, 95, 97, 98, 100],
    [1, 3, 4, 6, 10, 11, 13, 14, 15, 16, 17, 18, 19, 20, 21, 24, 25, 31, 32, 33, 34, 36, 37, 38, 39, 43, 46, 47,
     50, 52, 53, 54, 56, 57, 58, 60, 61, 63, 64, 65, 68, 69, 70, 72, 73, 76, 78, 79, 80, 81, 82, 83, 84, 85, 86,
     88, 90, 92, 93, 94, 97, 98, 99, 100],
    [1, 3, 4, 6, 10, 11, 13, 14, 15, 16, 17, 18, 19, 21, 24, 25, 31, 32, 33, 34, 36, 37, 38, 39, 43, 46, 47, 50,
     52, 53, 54, 56, 57, 60, 61, 63, 64, 65, 68, 69, 70, 72, 73, 76, 77, 78, 79, 81, 82, 83, 84, 85, 86, 88, 90,
     92, 93, 94, 97, 98, 99, 100],
    [1, 3, 4, 6, 9, 10, 11, 13, 14, 15, 16, 18, 19, 20, 21, 24, 25, 27, 31, 33, 36, 37, 38, 39, 42, 43, 46, 47, 49,
     50, 54, 56, 57, 58, 59, 62, 63, 64, 65, 67, 68, 69, 70, 72, 73, 75, 76, 78, 80, 81, 82, 84, 85, 86, 87, 88,
     92, 95, 97, 98, 99, 100],
    [1, 3, 4, 6, 9, 11, 14, 15, 16, 18, 19, 20, 21, 24, 25, 26, 27, 31, 36, 37, 38, 39, 42, 43, 46, 47, 48, 49, 50,
     54, 56, 57, 58, 59, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 75, 76, 78, 80, 81, 84, 85, 86, 87,
     88, 92, 95, 97, 98, 99],
    [1, 3, 6, 9, 13, 14, 16, 18, 19, 20, 21, 24, 25, 30, 31, 32, 34, 35, 36, 38, 39, 46, 47, 48, 49, 50, 52, 53,
     56, 57, 58, 60, 61, 62, 63, 64, 65, 66, 67, 68, 70, 71, 72, 76, 78, 79, 80, 81, 82, 83, 84, 85, 86, 88, 90,
     92, 93, 95, 97, 98, 100],
    [1, 3, 4, 6, 9, 11, 13, 14, 15, 16, 17, 18, 19, 20, 24, 25, 28, 31, 32, 34, 35, 36, 37, 38, 39, 46, 47, 49,
     50, 52, 53, 54, 57, 58, 61, 63, 64, 65, 66, 67, 68, 69, 70, 71, 73, 76, 79, 80, 81, 83, 85, 86, 88, 90, 92,
     93, 94, 95, 97, 98, 100],
    [1, 3, 4, 9, 10, 11, 13, 14, 15, 16, 19, 20, 21, 22, 24, 25, 32, 33, 34, 36, 37, 38, 39, 40, 41, 43, 47, 49,
     50, 52, 53, 54, 56, 57, 58, 60, 61, 63, 65, 67, 68, 69, 70, 73, 76, 78, 79, 80, 81, 82, 83, 85, 88, 90, 93,
     95, 97, 98, 99, 100],
    [1, 3, 4, 6, 10, 11, 13, 14, 15, 16, 17, 18, 24, 25, 31, 32, 33, 34, 35, 36, 37, 38, 39, 44, 45, 46, 47, 50,
     52, 53, 54, 56, 57, 60, 61, 63, 64, 65, 68, 69, 70, 72, 73, 76, 77, 78, 79, 82, 83, 85, 86, 88, 90, 92, 93,
     94, 98, 100],
    [1, 3, 4, 10, 11, 12, 13, 14, 15, 16, 17, 19, 21, 22, 24, 25, 32, 33, 34, 36, 37, 38, 39, 40, 41, 43, 47, 50,
     52, 54, 55, 56, 57, 60, 61, 63, 65, 68, 69, 70, 73, 76, 77, 78, 79, 81, 82, 83, 85, 88, 90, 93, 94, 97, 98,
     99, 100],
    [1, 3, 4, 6, 9, 10, 11, 13, 14, 15, 16, 17, 18, 20, 24, 25, 28, 29, 31, 33, 34, 35, 36, 37, 38, 39, 44, 46,
     47, 50, 52, 53, 54, 57, 58, 63, 64, 65, 67, 68, 69, 70, 73, 76, 79, 80, 81, 85, 86, 88, 90, 92, 93, 94, 98],
    [3, 6, 9, 13, 14, 16, 18, 20, 21, 24, 25, 30, 32, 34, 38, 39, 43, 46, 47, 48, 49, 50, 52, 53, 56, 57, 58, 60,
     61, 62, 63, 64, 66, 67, 70, 71, 72, 76, 78, 79, 80, 81, 82, 83, 84, 85, 86, 88, 89, 90, 92, 97, 99, 100],
    [2, 5, 6, 9, 13, 16, 20, 21, 24, 25, 32, 34, 38, 43, 46, 48, 49, 50, 52, 53, 56, 57, 58, 60, 61, 62, 63, 66,
     67, 70, 71, 76, 78, 79, 80, 81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 99, 100],
    [1, 3, 4, 9, 13, 14, 15, 16, 20, 21, 22, 23, 24, 25, 32, 34, 37, 39, 40, 43, 47, 48, 49, 50, 52, 53, 56, 57,
     58, 61, 66, 67, 69, 70, 71, 76, 79, 80, 81, 82, 83, 85, 88, 90, 93, 95, 98, 99, 100],
    [1, 3, 6, 7, 9, 13, 16, 17, 18, 20, 24, 25, 31, 34, 35, 36, 37, 38, 39, 44, 45, 46, 47, 48, 49, 50, 57, 58,
     63, 64, 65, 66, 67, 68, 69, 70, 71, 73, 79, 80, 81, 85, 86, 88, 90, 92, 94, 95, 98],
    [1, 3, 6, 7, 8, 9, 13, 16, 17, 18, 20, 24, 25, 31, 34, 35, 36, 37, 38, 39, 44, 45, 46, 47, 48, 49, 57, 58, 63,
     64, 65, 66, 67, 68, 69, 70, 71, 73, 79, 80, 81, 85, 86, 88, 90, 92, 95, 98],
    [2, 3, 5, 6, 10, 13, 14, 16, 17, 21, 24, 25, 32, 33, 34, 38, 43, 46, 50, 52, 53, 56, 57, 60, 61, 63, 70, 76,
     77, 78, 79, 81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 96, 99, 100],
    [1, 3, 4, 6, 8, 10, 13, 15, 16, 18, 24, 25, 31, 33, 34, 35, 36, 37, 38, 39, 46, 47, 51, 57, 63, 64, 65, 68,
     69, 70, 73, 77, 79, 85, 86, 88, 90, 92, 98]
]


# === 增强版指标收集器 ===
class MetricsCollector:
    def __init__(self):
        self.pso_start_time = None
        self.pso_end_time = None
        self.perfect_solutions_count = 0
        self.total_paths_count = 0
        self.final_fitness_scores = []
        self.pso_convergence_iterations = []
        self.pso_reset_counts = []
        self.path_execution_times = []

    def start_pso_phase(self):
        self.pso_start_time = time.time()

    def end_pso_phase(self):
        self.pso_end_time = time.time()

    def record_pso_result(self, fitness, is_perfect_match, convergence_iter=None, path_id=None, method='MI-PSO',
                          reset_count=0, execution_time=0):
        self.final_fitness_scores.append(fitness)
        self.total_paths_count += 1
        self.pso_reset_counts.append(reset_count)
        self.path_execution_times.append(execution_time)

        if is_perfect_match:
            self.perfect_solutions_count += 1
        if convergence_iter is not None:
            self.pso_convergence_iterations.append(convergence_iter)


# === 粒子类 ===
class Particle:
    def __init__(self):
        self.position = np.array([
            np.random.uniform(BOUNDS['light'][0], BOUNDS['light'][1]),
            np.random.uniform(BOUNDS['temp'][0], BOUNDS['temp'][1]),
            np.random.uniform(BOUNDS['moisture'][0], BOUNDS['moisture'][1])
        ])
        self.velocity = np.array([
            np.random.uniform(-5, 5),
            np.random.uniform(-3, 3),
            np.random.uniform(-5, 5)
        ])
        self.best_position = self.position.copy()
        self.best_fitness = 0
        self.fitness = 0


# === 单独的 MI-PSO 优化器类（已完全修复随机性重评估Bug） ===
class MIPSO:
    def __init__(self, target_path, swarm_size=20, max_iterations=3000, g2=10, th_cv=1.2):
        self.target_path = target_path
        self.swarm_size = swarm_size
        self.max_iterations = max_iterations
        self.g2 = g2
        self.th_cv = th_cv

        # 完全随机初始化
        self.particles = [Particle() for _ in range(swarm_size)]
        self.global_best_position = None
        self.global_best_fitness = 0
        self.global_best_triggered = set()  # 核心修复：记录最高分对应的真实触发路径
        self.reset_count = 0

        # 初始化适应度评估
        for particle in self.particles:
            fit, trig = self.evaluate(particle.position)
            particle.fitness = fit
            particle.best_fitness = fit
            particle.best_position = particle.position.copy()

            if fit > self.global_best_fitness:
                self.global_best_fitness = fit
                self.global_best_position = particle.position.copy()
                self.global_best_triggered = trig

    def evaluate(self, position):
        """同时返回适应度和实际触发的路径"""
        try:
            triggered = execute_Tr(position)
            if self.target_path.issubset(triggered):
                return 1.0, triggered
            intersection = len(triggered & self.target_path)
            union = len(triggered | self.target_path)
            return (intersection / union if union > 0 else 0.0), triggered
        except:
            return 0.0, set()

    def calculate_cv(self, scores):
        """实现公式 25：综合变异系数"""
        N = self.swarm_size
        mean_fit = np.mean(scores)
        if mean_fit == 0: return float('inf')

        var_term = np.sum((scores - mean_fit) ** 2) / N
        adj_diff = sum(abs(scores[2 * i + 1] - scores[2 * i]) for i in range(N // 2))
        adj_term = ((2 / N) * adj_diff) ** 2

        numerator = math.sqrt(var_term + adj_term)
        denominator = 2 * np.sum(scores)

        if denominator == 0: return float('inf')
        return numerator / denominator

    def flip_mutation(self, position, rho=0.1):
        """实现公式 27：变异反转操作"""
        new_pos = np.copy(position)
        new_pos[0] = BOUNDS['light'][0] + BOUNDS['light'][1] - position[0] + rho
        new_pos[1] = BOUNDS['temp'][0] + BOUNDS['temp'][1] - position[1] + rho
        new_pos[2] = BOUNDS['moisture'][0] + BOUNDS['moisture'][1] - position[2] + rho
        return clip_state(new_pos)

    def optimize(self):
        w = 0.7
        c1 = 1.5
        c2 = 1.5

        max_velocity = np.array([
            (BOUNDS['light'][1] - BOUNDS['light'][0]) * 0.2,
            (BOUNDS['temp'][1] - BOUNDS['temp'][0]) * 0.2,
            (BOUNDS['moisture'][1] - BOUNDS['moisture'][0]) * 0.2
        ])

        c = 0
        early_stop = False
        converged_at = self.max_iterations

        for iteration in range(self.max_iterations):
            scores = np.array([p.fitness for p in self.particles])
            c += 1

            # MI-PSO 特有的停滞检测与变异 (对应 Algorithm 4 Line 20-23)
            if c == self.g2:
                cv = self.calculate_cv(scores)
                if cv <= self.th_cv:
                    self.reset_count += 1
                    l = int(round(self.swarm_size * (1 - cv)))
                    l = min(max(l, 1), self.swarm_size)

                    selected_indices = np.random.choice(self.swarm_size, l, replace=False)
                    for idx in selected_indices:
                        new_pos = self.flip_mutation(self.particles[idx].position)
                        new_fit, new_trig = self.evaluate(new_pos)

                        if new_fit > self.particles[idx].fitness:
                            self.particles[idx].position = new_pos
                            self.particles[idx].fitness = new_fit
                            if new_fit > self.particles[idx].best_fitness:
                                self.particles[idx].best_fitness = new_fit
                                self.particles[idx].best_position = new_pos.copy()
                            if new_fit > self.global_best_fitness:
                                self.global_best_fitness = new_fit
                                self.global_best_position = new_pos.copy()
                                self.global_best_triggered = new_trig  # 更新最优路径
                c = 0

                # 标准 PSO 速度与位置更新
            for particle in self.particles:
                r1, r2 = np.random.random(3), np.random.random(3)
                particle.velocity = (w * particle.velocity +
                                     c1 * r1 * (particle.best_position - particle.position) +
                                     c2 * r2 * (self.global_best_position - particle.position))
                particle.velocity = np.clip(particle.velocity, -max_velocity, max_velocity)

                particle.position += particle.velocity
                particle.position = clip_state(particle.position)

                fit, trig = self.evaluate(particle.position)
                particle.fitness = fit

                if fit > particle.best_fitness:
                    particle.best_fitness = fit
                    particle.best_position = particle.position.copy()
                if fit > self.global_best_fitness:
                    self.global_best_fitness = fit
                    self.global_best_position = particle.position.copy()
                    self.global_best_triggered = trig  # 更新最优路径

            if self.global_best_fitness >= 1.0:
                converged_at = iteration + 1
                early_stop = True
                break

        return self.global_best_position, self.global_best_fitness, self.global_best_triggered, converged_at, early_stop, self.reset_count


# === Excel导出函数 ===
def export_multiple_runs_to_excel(all_run_results, all_run_metrics, num_runs, filename=None):
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Isolated_MIPSO_{num_runs}Runs_{timestamp}.xlsx"

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    border = Border(
        left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 工作表1: 运行汇总
    ws1 = wb.active
    ws1.title = "运行汇总"
    ws1.sheet_view.showGridLines = False

    headers = ["运行次数", "成功率", "成功数量", "平均适应度", "平均迭代次数", "运行时间(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, border, center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, (results, run_metrics) in enumerate(zip(all_run_results, all_run_metrics), start=1):
        success_count = sum(1 for r in results if r['perfect_match'])
        success_rate = (success_count / len(targetPaths)) * 100
        avg_fitness = np.mean([r['fitness'] for r in results])

        iterations_list = [r['convergence_iteration'] if r.get('convergence_iteration') is not None else 10000 for r in
                           results]
        avg_iterations = np.mean(iterations_list)
        total_time = run_metrics.pso_end_time - run_metrics.pso_start_time if run_metrics.pso_end_time else 0

        row_data = [f"运行 {run_idx}", f"{success_rate:.1f}%", f"{success_count}/{len(targetPaths)}",
                    f"{avg_fitness:.4f}", f"{avg_iterations:.1f}", f"{total_time:.2f}"]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border, cell.alignment = border, center_align
            if run_idx % 2 == 0: cell.fill = alternate_fill
            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    # 工作表2: 路径统计
    ws2 = wb.create_sheet(title="路径统计")
    ws2.sheet_view.showGridLines = False
    headers2 = ["路径编号", "成功次数", "成功率", "平均适应度", "平均迭代次数", "最小迭代", "最大迭代"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, border, center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(targetPaths)):
        success_count = sum(1 for results in all_run_results if results[path_idx]['perfect_match'])
        success_rate = (success_count / num_runs) * 100
        avg_fitness = np.mean([results[path_idx]['fitness'] for results in all_run_results])

        iterations_list = [results[path_idx]['convergence_iteration'] if results[path_idx].get(
            'convergence_iteration') is not None else 10000 for results in all_run_results]
        avg_iterations = np.mean(iterations_list)

        row_data = [f"路径 {path_idx + 1}", f"{success_count}/{num_runs}", f"{success_rate:.1f}%",
                    f"{avg_fitness:.4f}", f"{avg_iterations:.1f}", np.min(iterations_list), np.max(iterations_list)]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border, cell.alignment = border, center_align
            if (path_idx + 1) % 2 == 0: cell.fill = alternate_fill
            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    # 工作表3: 最佳粒子详情
    ws3 = wb.create_sheet(title="最佳粒子详情")
    ws3.sheet_view.showGridLines = False
    headers3 = ["路径", "运行", "最佳粒子(light,temp,moisture)", "适应度", "迭代次数", "求解方法", "生成路径"]
    col_widths3 = [10, 10, 25, 12, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, border, center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(targetPaths)):
        for run_idx, results in enumerate(all_run_results, start=1):
            result = results[path_idx]
            particle_str = f"({int(result['best_position'][0])}, {int(result['best_position'][1])}, {int(result['best_position'][2])})"
            path_str = str(sorted(list(result['triggered'])))

            convergence_iter = result['convergence_iteration'] if result.get(
                'convergence_iteration') is not None else 10000

            row_data = [f"路径{path_idx + 1}", f"运行{run_idx}", particle_str, f"{result['fitness']:.4f}",
                        convergence_iter if convergence_iter < 10000 else "-", "MI-PSO", path_str]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = left_align if col == 7 else center_align

                if result['fitness'] == 1.0:
                    cell.fill = success_fill
                elif result['fitness'] < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill
            row_idx += 1

    # 保存
    output_dir = os.path.join(os.getcwd(), "results")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"\n✓ 独立 MI-PSO 消融实验结果已导出到: {filepath}")
    return filepath


def run_single_mipso_experiment(run_num):
    print(f"\n{'=' * 50}")
    print(f"开始第 {run_num} 次运行 (独立 MI-PSO)")
    print(f"{'=' * 50}")

    run_metrics = MetricsCollector()
    run_metrics.start_pso_phase()

    pso_results = []

    for i, target_path in enumerate(targetPaths):
        path_start_time = time.time()

        mipso = MIPSO(target_path, swarm_size=20, max_iterations=3000)
        # 直接接收并使用最佳寻优过程中真实触发的 best_trig，坚决不重新 evaluate
        best_pos, best_fit, best_trig, converged_at, early_stop, reset_count = mipso.optimize()

        path_execution_time = time.time() - path_start_time

        # 使用真实的最高分数判断是否完美
        is_perfect = (best_fit >= 1.0)

        pso_results.append({
            'target_path': target_path,
            'best_position': best_pos,
            'fitness': best_fit,
            'triggered': best_trig,
            'perfect_match': is_perfect,
            'method': 'MI-PSO',
            'convergence_iteration': converged_at,
            'early_stopped': early_stop,
            'reset_count': reset_count
        })

        run_metrics.record_pso_result(
            fitness=best_fit,
            is_perfect_match=is_perfect,
            convergence_iter=converged_at if early_stop else None,
            path_id=i + 1,
            method='MI-PSO',
            reset_count=reset_count,
            execution_time=path_execution_time
        )

        status = "✓完美(MI-PSO)" if is_perfect else f"○部分({best_fit:.3f})"
        print(f"  路径{i + 1}: {status} | 变异触发次数: {reset_count} | 耗时 {path_execution_time:.2f}s")

    run_metrics.end_pso_phase()

    success_count = sum(1 for r in pso_results if r['perfect_match'])
    success_rate = (success_count / len(targetPaths)) * 100
    pso_time = run_metrics.pso_end_time - run_metrics.pso_start_time

    print(f"\n第{run_num}次运行完成: 成功率 {success_rate:.1f}% | 耗时 {pso_time:.2f}秒")
    return pso_results, run_metrics


def run_multiple_mipso_experiments(num_runs):
    all_run_results = []
    all_run_metrics = []
    total_start = time.time()

    for run_num in range(1, num_runs + 1):
        results, metrics = run_single_mipso_experiment(run_num)
        all_run_results.append(results)
        all_run_metrics.append(metrics)

    total_time = time.time() - total_start
    print(f"\n{'=' * 50}\n全部{num_runs}次独立 MI-PSO 运行完成! 总耗时: {total_time:.2f}秒\n{'=' * 50}")

    return all_run_results, all_run_metrics


if __name__ == "__main__":
    if len(sys.argv) > 1:
        try:
            NUM_RUNS = int(sys.argv[1])
        except ValueError:
            pass

    print("=" * 70)
    print(" 独立 MI-PSO 消融实验 (无 DQN 介入，完美修复随机性 Bug)")
    print("=" * 70)

    all_run_results, all_run_metrics = run_multiple_mipso_experiments(NUM_RUNS)
    export_multiple_runs_to_excel(all_run_results, all_run_metrics, NUM_RUNS)