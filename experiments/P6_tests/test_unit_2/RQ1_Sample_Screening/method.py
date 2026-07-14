import os
import random
from collections import deque
import numpy as np
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# --- 全局状态范围配置 ---
STATE_MIN_X, STATE_MAX_X = 1, 1000
STATE_MIN_Y, STATE_MAX_Y = 1, 1000
STATE_MIN_Z, STATE_MAX_Z = 1, 600

def generate_input():
    return [
        random.randint(STATE_MIN_X, STATE_MAX_X),
        random.randint(STATE_MIN_Y, STATE_MAX_Y),
        random.randint(STATE_MIN_Z, STATE_MAX_Z)
    ]


def execute_Tr(x, y, z):
    triggered = set()

    # 固定参数
    energy_price = 0.15
    time_of_day = 12
    temp = 25
    humidity = 60
    energy_trend = "stable"
    weather_forecast = "sunny"

    # 分支1-10: 能量价格相关规则
    if (energy_price > 0.10 and time_of_day == 12 and (x + temp) > 500) != (energy_price > 0.10 and time_of_day == 12 and (x + temp) > 800):
        triggered.add(1)
    if (energy_price > 0.10 and time_of_day == 12 and (x + temp) > 1000) != (energy_price > 0.10 and time_of_day == 12 and (x + temp) > 900):
        triggered.add(2)
    if (energy_price < 0.20 and (x + z) < 800) != (energy_price < 0.20 and (x + z) < 1200):
        triggered.add(3)
    if (energy_price < 0.20 and (x + z) < 600) != (energy_price < 0.20 and (x + z) < 1000):
        triggered.add(4)
    if (energy_price > 0.10 and (x * temp) > 20000) != (energy_price > 0.10 and (x * temp) > 30000):
        triggered.add(5)
    if (energy_price > 0.10 and (x * temp) > 15000) != (energy_price > 0.10 and (x * temp) > 25000):
        triggered.add(6)
    if (energy_price < 0.20 and (y + humidity) < 100) != (energy_price < 0.20 and (y + humidity) < 150):
        triggered.add(7)
    if (energy_price < 0.20 and (y + humidity) < 80) != (energy_price < 0.20 and (y + humidity) < 120):
        triggered.add(8)
    if (energy_price > 0.10 and z > 300) != (energy_price > 0.10 and z > 450):
        triggered.add(9)
    if (energy_price > 0.10 and z > 200) != (energy_price > 0.10 and z > 350):
        triggered.add(10)

    # 分支11-19: 时间段相关规则
    if (time_of_day == 12 and x < 500 and y > 200) != (time_of_day == 12 and x < 700 and y > 200):
        triggered.add(11)
    if (time_of_day == 12 and x < 400 and z < 300) != (time_of_day == 12 and x < 600 and z < 300):
        triggered.add(12)
    if (temp + humidity > 20 and y < 500) != (temp + humidity < 20 and y < 500):
        triggered.add(13)
    if (energy_price < 0.20 and (z + x) < 800 and y > 100) != (energy_price < 0.20 and (z + x) < 1200 and y > 100):
        triggered.add(14)
    if (energy_price < 0.20 and (z + x) < 600 and y < 800) != (energy_price < 0.20 and (z + x) < 1000 and y < 800):
        triggered.add(15)
    if (temp < 30 and x > 500 and z > 200) != (temp < 30 and x < 500 and z > 200):
        triggered.add(16)
    if (temp < 28 and x > 400 and y > 300) != (temp < 28 and x > 400 and z > 300):
        triggered.add(17)
    if (time_of_day == 12 and x > 600 and z < 400) != (time_of_day == 12 and x > 800 and z < 400):
        triggered.add(18)
    if (time_of_day == 12 and x > 500 and y + z > 600) != (time_of_day == 12 and x > 750 and y + z > 600):
        triggered.add(19)

    # 分支20-29: 更多组合规则
    if (time_of_day == 12 and z < 300 and x > 200) != (time_of_day == 12 and z < 450 and x > 200):
        triggered.add(20)
    if (time_of_day == 12 and z < 250 and y > 300) != (time_of_day == 12 and z < 400 and y > 300):
        triggered.add(21)
    if (energy_price > 0.10 and x < 400 and y + z > 500) != (energy_price > 0.10 and x < 600 and y + z > 500):
        triggered.add(22)
    if (energy_price > 0.10 and x < 300 and z > 200) != (energy_price > 0.10 and x < 500 and z > 200):
        triggered.add(23)
    if (time_of_day == 12 and temp > 20 and x > 400) != (time_of_day == 12 and temp > 30 and x > 400):
        triggered.add(24)
    if (time_of_day == 12 and temp > 18 and y < 600) != (time_of_day == 12 and temp > 28 and y < 600):
        triggered.add(25)
    if (humidity > 50 and y > 500 and x < 800) != (humidity > 70 and y > 500 and x < 800):
        triggered.add(26)
    if (humidity > 45 and y > 450 and z < 400) != (humidity > 65 and y > 450 and z < 400):
        triggered.add(27)
    if (time_of_day == 12 and y < 200 and x + z > 500) != (time_of_day == 12 and y < 300 and x + z > 500):
        triggered.add(28)
    if (time_of_day == 12 and y < 150 and z > 100) != (time_of_day == 12 and y < 250 and z > 100):
        triggered.add(29)

    # 分支30-37: 更多规则
    if (energy_price > 0.10 and x > 700) != (energy_price > 0.10 and x > 850):
        triggered.add(30)
    if (energy_price > 0.10 and x > 600) != (energy_price > 0.10 and x > 750):
        triggered.add(31)
    if (temp < 28 and y < 250) != (temp < 28 and y < 450):
        triggered.add(32)
    if (energy_price > 0.10 and z > 350) != (energy_price > 0.10 and x > 350):
        triggered.add(33)
    if (humidity < 70 and y < 400) != (humidity < 30 and y < 400):
        triggered.add(34)
    if (humidity < 65 and y < 350) != (humidity < 65 and x < 350):
        triggered.add(35)
    if (time_of_day == 12 and (x + temp) < 600) != (time_of_day == 12 and (x + temp) < 800):
        triggered.add(36)
    if (time_of_day == 12 and (x + temp) < 500) != (time_of_day == 12 and (x + temp) < 700):
        triggered.add(37)

    # 分支38-53: 复杂表达式
    if (energy_price > 0.10 and (x * 0.15) > 80 and temp > 20) != (energy_price > 0.10 and (x * 0.15) > 150 and temp > 20):
        triggered.add(38)
    if (energy_price > 0.10 and (x * 0.15) > 60 and temp > 20) != (energy_price > 0.10 and (x * 0.15) > 120 and temp > 20):
        triggered.add(39)
    if (energy_price < 0.20 and (z * 0.5) > 100 and x < 600) != (energy_price < 0.20 and (z * 0.5) > 200 and x < 600):
        triggered.add(40)
    if (energy_price < 0.20 and (z * 0.5) > 80 and x < 600) != (energy_price < 0.20 and (z * 0.5) > 180 and x < 600):
        triggered.add(41)
    if (temp + humidity > 80 and y > 400) != (temp + humidity > 100 and y > 400):
        triggered.add(42)
    if (temp + humidity > 70 and y > 350) != (temp + humidity > 90 and y > 350):
        triggered.add(43)
    if (energy_price < 0.20 and (y + z) < 400) != (energy_price < 0.20 and (y + z) < 600):
        triggered.add(44)
    if (energy_price < 0.20 and (y + z) < 350) != (energy_price < 0.20 and (y + z) < 550):
        triggered.add(45)
    if (time_of_day == 12 and (x * temp) > 15000) != (time_of_day == 12 and (x * temp) > 25000):
        triggered.add(46)
    if (time_of_day == 12 and (x * temp) > 12000) != (time_of_day == 12 and (x * temp) > 20000):
        triggered.add(47)
    if (time_of_day == 12 and (z + temp) < 400) != (time_of_day == 12 and (z + temp) < 600):
        triggered.add(48)
    if (time_of_day == 12 and (z + temp) < 350) != (time_of_day == 12 and (z + temp) < 550):
        triggered.add(49)
    if (energy_price > 0.10 and (humidity * 0.15) > 8 and y > 70) != (energy_price > 0.10 and (humidity * 0.15) > 8 and y > 870):
        triggered.add(50)
    if (energy_price < 0.20 and (x * y * 0.01) < 200 and z < 300) != (energy_price < 0.20 and (x * y * 0.01) < 400 and z < 300):
        triggered.add(51)
    if (time_of_day == 12 and (humidity + y) * 0.5 < 150) != (time_of_day == 12 and (humidity + y) * 0.5 < 250):
        triggered.add(52)
    if (time_of_day == 12 and (humidity + y) * 0.5 < 120) != (time_of_day == 12 and (humidity + y) * 0.5 < 220):
        triggered.add(53)

    # 分支54-63: energy_trend相关
    if (energy_trend == "stable" and x < 500 and y > 200) != (energy_trend == "stable" and x < 700 and y > 200):
        triggered.add(54)
    if (energy_trend == "stable" and x < 400 and z < 300) != (energy_trend == "stable" and x < 600 and z < 300):
        triggered.add(55)
    if (energy_price > 0.10 and temp > 20 and y > 300) != (energy_price > 0.10 and temp > 30 and y > 300):
        triggered.add(56)
    if (energy_price > 0.10 and temp > 18 and z < 400) != (energy_price > 0.10 and temp > 28 and z < 400):
        triggered.add(57)
    if (energy_trend == "stable" and z < 300 and x + y > 600) != (energy_trend == "stable" and z < 450 and x + y > 600):
        triggered.add(58)
    if (energy_trend == "stable" and z < 250 and y > 400) != (energy_trend == "stable" and z < 400 and y > 400):
        triggered.add(59)
    if (energy_trend == "stable" and y < 300 and x > 500) != (energy_trend == "stable" and y < 450 and x > 500):
        triggered.add(60)
    if (energy_trend == "stable" and y < 250 and z > 200) != (energy_trend == "stable" and y < 400 and z > 200):
        triggered.add(61)
    if (humidity > 50 and x > 600 and y + z > 700) != (humidity > 70 and x > 600 and y + z > 700):
        triggered.add(62)
    if (humidity > 45 and x > 550 and z < 400) != (humidity > 65 and x > 550 and z < 400):
        triggered.add(63)

    # 分支64-71: 更多组合
    if (energy_trend == "stable" and time_of_day == 12 and x > 600) != (energy_trend == "stable" and time_of_day == 12 and x > 800):
        triggered.add(64)
    if (energy_trend == "stable" and time_of_day == 12 and x > 500) != (energy_trend == "stable" and time_of_day == 12 and x > 750):
        triggered.add(65)
    if (energy_price < 0.20 and z > 350) != (energy_price < 0.20 and z > 450):
        triggered.add(66)
    if (energy_price < 0.20 and z > 300) != (energy_price < 0.20 and z > 400):
        triggered.add(67)
    if (energy_trend == "stable" and (x + temp) < 600) != (energy_trend == "stable" and (x + temp) < 800):
        triggered.add(68)
    if (energy_trend == "stable" and (x + temp) < 500) != (energy_trend == "stable" and (x + temp) < 700):
        triggered.add(69)
    if (humidity < 70 and z < 400) != (humidity < 70 and x < 400):
        triggered.add(70)
    if (humidity < 65 and z < 350) != (humidity < 65 and z < 750):
        triggered.add(71)

    # 分支72-81: weather_forecast相关
    if (weather_forecast == "sunny" and x < 600 and y > 300) != (weather_forecast == "sunny" and x < 800 and y > 300):
        triggered.add(72)
    if (weather_forecast == "sunny" and x < 500 and z < 400) != (weather_forecast == "sunny" and x < 700 and z < 400):
        triggered.add(73)
    if (weather_forecast == "sunny" and y < 500 and x > 200) != (weather_forecast == "sunny" and y < 700 and x > 200):
        triggered.add(74)
    if (weather_forecast == "sunny" and y < 400 and z > 250) != (weather_forecast == "sunny" and y < 600 and z > 250):
        triggered.add(75)
    if (weather_forecast == "sunny" and humidity > 50 and x + y > 800) != (weather_forecast == "sunny" and humidity > 70 and x + y > 800):
        triggered.add(76)
    if (weather_forecast == "sunny" and humidity > 45 and z < 300) != (weather_forecast == "sunny" and humidity > 65 and z < 300):
        triggered.add(77)
    if (weather_forecast == "sunny" and temp > 20 and x * y > 200000) != (weather_forecast == "sunny" and temp > 30 and x * y > 200000):
        triggered.add(78)
    if (weather_forecast == "sunny" and temp > 18 and y + z < 800) != (weather_forecast == "sunny" and temp > 28 and y + z < 800):
        triggered.add(79)
    if (weather_forecast == "sunny" and z < 300 and x > 400) != (weather_forecast == "sunny" and z < 450 and x > 400):
        triggered.add(80)
    if (weather_forecast == "sunny" and z < 250 and y < 600) != (weather_forecast == "sunny" and z < 400 and y < 600):
        triggered.add(81)

    # 分支82-85: 高级组合
    if (weather_forecast == "sunny" and time_of_day == 12 and y < 300 and x > 400) != (weather_forecast == "sunny" and time_of_day == 12 and y < 500 and x > 400):
        triggered.add(82)
    if (weather_forecast == "sunny" and time_of_day == 12 and y < 250 and z > 200) != (weather_forecast == "sunny" and time_of_day == 12 and y < 450 and z > 200):
        triggered.add(83)
    if (weather_forecast == "sunny" and energy_price > 0.10 and x + y > 1000) != (weather_forecast == "sunny" and energy_price > 0.20 and x + y > 1000):
        triggered.add(84)
    if (weather_forecast == "sunny" and energy_price > 0.05 and z < 400) != (weather_forecast == "sunny" and energy_price > 0.18 and z < 400):
        triggered.add(85)

    # 分支86-100: 最复杂的规则
    if (energy_price * 10 + time_of_day * 0.5 > 7 and x > 600 and z > 200) != (energy_price * 10 + time_of_day * 0.5 > 11 and x > 600 and z > 200):
        triggered.add(86)
    if (energy_price * temp > 3 and humidity > 50 and x * y > 300000) != (energy_price * temp > 5 and humidity > 50 and x * y > 300000):
        triggered.add(87)
    if (energy_price * temp > 2.5 and humidity > 50 and z < 400) != (energy_price * temp > 4.5 and humidity > 50 and z < 400):
        triggered.add(88)
    if ((1 - energy_price * 2) * (x + z) > 800 and y < 400) != ((1 - energy_price * 2) * (x + z) > 1200 and y < 400):
        triggered.add(89)
    if ((1 - energy_price * 2) * (x + z) > 600 and y < 400) != ((1 - energy_price * 2) * (x + z) > 1000 and y < 400):
        triggered.add(90)
    if (time_of_day == 12 and x * temp * 0.1 > 2000 and y + z > 600) != (time_of_day == 12 and x * temp * 0.1 > 3000 and y + z > 600):
        triggered.add(91)
    if (time_of_day == 12 and x * temp * 0.1 > 1500 and z < 500) != (time_of_day == 12 and x * temp * 0.1 > 2500 and z < 500):
        triggered.add(92)
    if (time_of_day == 12 and (z + y) * 0.8 < 400 and x > 300) != (time_of_day == 12 and (z + y) * 0.8 < 600 and x > 300):
        triggered.add(93)
    if (time_of_day == 12 and (z + y) * 0.8 < 350 and x < 700) != (time_of_day == 12 and (z + y) * 0.8 < 550 and x < 700):
        triggered.add(94)
    if (energy_price * 10 + humidity * 0.1 > 7 and temp > 20 and z > x / 2) != (energy_price * 10 + humidity * 0.1 > 11 and temp > 20 and z > x / 2):
        triggered.add(95)
    if ((1 - energy_price * 3) * (x + temp) > 500 and z < 300 and x + y > 800) != ((1 - energy_price * 3) * (x + temp) > 800 and z < 300 and x + y > 800):
        triggered.add(96)
    if (energy_price > 0.10 and abs(time_of_day - 12) < 1 and (x + temp) > 600 and y * z > 100000) != (energy_price > 0.10 and abs(time_of_day - 12) < 1 and (x + temp) > 900 and y * z > 100000):
        triggered.add(97)
    if (energy_price > 0.10 and abs(time_of_day - 12) < 1 and (x + temp) > 500 and y + z < 1000) != (energy_price > 0.10 and abs(time_of_day - 12) < 1 and (x + temp) > 800 and y + z < 1000):
        triggered.add(98)
    if (24 - time_of_day > 8 and (y + humidity) < 400 and x > z) != (24 - time_of_day > 8 and (y + humidity) < 600 and x > z):
        triggered.add(99)
    if (24 - time_of_day > 8 and (y + humidity) < 350 and x + y + z > 1000) != (24 - time_of_day > 8 and (y + humidity) < 550 and x + y + z > 1000):
        triggered.add(100)

    return triggered


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0


# Fixed: Convert all to lists for consistency
TARGET_PATHS = [
    [1, 3, 6, 9, 11, 13, 14, 16, 18, 19, 20, 21, 24, 25, 31, 32, 34, 36, 37, 38, 39, 43, 46, 47, 48, 49, 50,
     52, 53, 54, 56, 57, 58, 60, 61, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 76, 78, 79, 80, 81, 82,
     83, 85, 86, 88, 90, 92, 93, 95, 97, 98, 99, 100],
    [1, 3, 4, 6, 9, 11, 13, 14, 15, 16, 18, 19, 20, 21, 24, 25, 31, 32, 34, 36, 37, 38, 39, 43, 46, 47, 49, 50, 52,
     54, 56, 57, 58, 60, 61, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 76, 78, 79, 80, 81, 82, 83, 84, 85,
     86, 88, 90, 92, 95, 97, 98, 99, 100],
    [1, 3, 4, 6, 9, 10, 11, 13, 14, 15, 16, 18, 19, 20, 21, 24, 25, 31, 32, 33, 36, 37, 38, 39, 42, 43, 46, 47, 50,
     52, 54, 56, 57, 58, 59, 60, 62, 63, 64, 65, 67, 68, 69, 70, 72, 73, 75, 76, 78, 79, 80, 81, 82, 83, 84, 85,
     86, 88, 92, 93, 95, 97, 98, 99, 100],
    [1, 3, 4, 6, 9, 10, 11, 13, 14, 15, 16, 18, 19, 20, 21, 24, 25, 31, 32, 33, 34, 35, 36, 37, 38, 39, 46, 47, 49,
     50, 52, 53, 54, 56, 57, 58, 60, 61, 63, 64, 65, 67, 68, 69, 70, 72, 73, 76, 78, 79, 80, 81, 82, 83, 85, 86,
     88, 90, 92, 93, 94, 95, 97, 98, 100],
    [1, 3, 4, 6, 10, 11, 13, 14, 15, 16, 17, 18, 19, 20, 21, 24, 25, 31, 32, 33, 34, 36, 37, 38, 39, 43, 46, 47,
     50, 52, 53, 54, 56, 57, 58, 60, 61, 63, 64, 65, 68, 69, 70, 72, 73, 76, 78, 79, 80, 81, 82, 83, 84, 85, 86,
     88, 90, 92, 93, 94, 97, 98, 99, 100],
    [1, 3, 4, 6, 10, 11, 13, 14, 15, 16, 17, 18, 19, 21, 24, 25, 31, 32, 33, 34, 36, 37, 38, 39, 43, 46, 47, 50,
     52, 53, 54, 56, 57, 60, 61, 63, 64, 65, 68, 69, 70, 72, 73, 76, 77, 78, 79, 81, 82, 83, 84, 85, 86, 88, 90,
     92, 93, 94, 97, 98, 99, 100],
    [1, 3, 4, 6, 9, 10, 11, 13, 14, 15, 16, 18, 19, 20, 21, 24, 25, 27, 31, 33, 36, 37, 38, 39, 42, 43, 46, 47, 49,
     50, 54, 56, 57, 58, 59, 62, 63, 64, 65, 67, 68, 69, 70, 72, 73, 75, 76, 78, 80, 81, 82, 84, 85, 86, 87, 88,
     92, 95, 97, 98, 99, 100],
    [1, 3, 4, 6, 9, 11, 14, 15, 16, 18, 19, 20, 21, 24, 25, 26, 27, 31, 36, 37, 38, 39, 42, 43, 46, 47, 48, 49, 50,
     54, 56, 57, 58, 59, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 75, 76, 78, 80, 81, 84, 85, 86, 87,
     88, 92, 95, 97, 98, 99],
    [1, 3, 6, 9, 13, 14, 16, 18, 19, 20, 21, 24, 25, 30, 31, 32, 34, 35, 36, 38, 39, 46, 47, 48, 49, 50, 52, 53,
     56, 57, 58, 60, 61, 62, 63, 64, 65, 66, 67, 68, 70, 71, 72, 76, 78, 79, 80, 81, 82, 83, 84, 85, 86, 88, 90,
     92, 93, 95, 97, 98, 100],
    [1, 3, 4, 6, 9, 11, 13, 14, 15, 16, 17, 18, 19, 20, 24, 25, 28, 31, 32, 34, 35, 36, 37, 38, 39, 46, 47, 49,
     50, 52, 53, 54, 57, 58, 61, 63, 64, 65, 66, 67, 68, 69, 70, 71, 73, 76, 79, 80, 81, 83, 85, 86, 88, 90, 92,
     93, 94, 95, 97, 98, 100],
    [1, 3, 4, 9, 10, 11, 13, 14, 15, 16, 19, 20, 21, 22, 24, 25, 32, 33, 34, 36, 37, 38, 39, 40, 41, 43, 47, 49,
     50, 52, 53, 54, 56, 57, 58, 60, 61, 63, 65, 67, 68, 69, 70, 73, 76, 78, 79, 80, 81, 82, 83, 85, 88, 90, 93,
     95, 97, 98, 99, 100],
    [1, 3, 4, 6, 10, 11, 13, 14, 15, 16, 17, 18, 24, 25, 31, 32, 33, 34, 35, 36, 37, 38, 39, 44, 45, 46, 47, 50,
     52, 53, 54, 56, 57, 60, 61, 63, 64, 65, 68, 69, 70, 72, 73, 76, 77, 78, 79, 82, 83, 85, 86, 88, 90, 92, 93,
     94, 98, 100],
    [1, 3, 4, 10, 11, 12, 13, 14, 15, 16, 17, 19, 21, 22, 24, 25, 32, 33, 34, 36, 37, 38, 39, 40, 41, 43, 47, 50,
     52, 54, 55, 56, 57, 60, 61, 63, 65, 68, 69, 70, 73, 76, 77, 78, 79, 81, 82, 83, 85, 88, 90, 93, 94, 97, 98,
     99, 100],
    [1, 3, 4, 6, 9, 10, 11, 13, 14, 15, 16, 17, 18, 20, 24, 25, 28, 29, 31, 33, 34, 35, 36, 37, 38, 39, 44, 46,
     47, 50, 52, 53, 54, 57, 58, 63, 64, 65, 67, 68, 69, 70, 73, 76, 79, 80, 81, 85, 86, 88, 90, 92, 93, 94, 98],
    [3, 6, 9, 13, 14, 16, 18, 20, 21, 24, 25, 30, 32, 34, 38, 39, 43, 46, 47, 48, 49, 50, 52, 53, 56, 57, 58, 60,
     61, 62, 63, 64, 66, 67, 70, 71, 72, 76, 78, 79, 80, 81, 82, 83, 84, 85, 86, 88, 89, 90, 92, 97, 99, 100],
    [2, 5, 6, 9, 13, 16, 20, 21, 24, 25, 32, 34, 38, 43, 46, 48, 49, 50, 52, 53, 56, 57, 58, 60, 61, 62, 63, 66,
     67, 70, 71, 76, 78, 79, 80, 81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 99, 100],
    [1, 3, 4, 9, 13, 14, 15, 16, 20, 21, 22, 23, 24, 25, 32, 34, 37, 39, 40, 43, 47, 48, 49, 50, 52, 53, 56, 57,
     58, 61, 66, 67, 69, 70, 71, 76, 79, 80, 81, 82, 83, 85, 88, 90, 93, 95, 98, 99, 100],
    [1, 3, 6, 7, 9, 13, 16, 17, 18, 20, 24, 25, 31, 34, 35, 36, 37, 38, 39, 44, 45, 46, 47, 48, 49, 50, 57, 58,
     63, 64, 65, 66, 67, 68, 69, 70, 71, 73, 79, 80, 81, 85, 86, 88, 90, 92, 94, 95, 98],
    [1, 3, 6, 7, 8, 9, 13, 16, 17, 18, 20, 24, 25, 31, 34, 35, 36, 37, 38, 39, 44, 45, 46, 47, 48, 49, 57, 58, 63,
     64, 65, 66, 67, 68, 69, 70, 71, 73, 79, 80, 81, 85, 86, 88, 90, 92, 95, 98],
    [2, 3, 5, 6, 10, 13, 14, 16, 17, 21, 24, 25, 32, 33, 34, 38, 43, 46, 50, 52, 53, 56, 57, 60, 61, 63, 70, 76,
     77, 78, 79, 81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 96, 99, 100],
    [1, 3, 4, 6, 8, 10, 13, 15, 16, 18, 24, 25, 31, 33, 34, 35, 36, 37, 38, 39, 46, 47, 51, 57, 63, 64, 65, 68,
     69, 70, 73, 77, 79, 85, 86, 88, 90, 92, 98]
]


class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 2000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(TARGET_PATHS)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }


def compute_robustness(state, path):
    base = execute_Tr(state[0], state[1], state[2])
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                    np.clip(state[0] + dx, STATE_MIN_X, STATE_MAX_X),
                    np.clip(state[1] + dy, STATE_MIN_Y, STATE_MAX_Y),
                    np.clip(state[2] + dz, STATE_MIN_Z, STATE_MAX_Z)
                ])
                n_trig = execute_Tr(neighbor[0], neighbor[1], neighbor[2])
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = TARGET_PATHS[target_path_idx]
    samples = []
    attempts = 0

    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(STATE_MIN_X, STATE_MAX_X),
            random.randint(STATE_MIN_Y, STATE_MAX_Y),
            random.randint(STATE_MIN_Z, STATE_MAX_Z)
        ])
        triggered = execute_Tr(state[0], state[1], state[2])

        if not triggered:
            continue

        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)

        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }

        samples.append(sample_data)

    return samples


def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample[
                'robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)

        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]

    return selected_samples


def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = TARGET_PATHS[target_path_idx]
        samples = []
        attempts = 0

        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(STATE_MIN_X, STATE_MAX_X),
                random.randint(STATE_MIN_Y, STATE_MAX_Y),
                random.randint(STATE_MIN_Z, STATE_MAX_Z)
            ])
            triggered = execute_Tr(state[0], state[1], state[2])

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)

            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)

        random.shuffle(samples)
        return samples[:config.top_k_samples]

    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")

        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)


def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}

    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates

    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}

        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config,
                                                         shared_candidates[path_idx])

            strategy_results[path_idx] = samples

        results[strategy_name] = strategy_results

    return results


def analyze_fitness_values(results, config):
    analysis_results = {}

    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []

        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])

        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }

        all_scores = []

        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])

        scores_array = np.array(all_scores)

        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)

        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })

        analysis_results[strategy_name] = analysis

    return analysis_results


def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []

    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)

    df = pd.DataFrame(df_data)
    return df, analysis_results


def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []

    print(f"Starting {num_runs} experiments...")

    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")

        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)

        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)

        print(f"Experiment {run_idx + 1} completed.")

    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df


def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')

    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width

    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []

    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]

        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)

    stats_df = pd.DataFrame(stats_data)

    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Results saved to: {output_path}")


def main():
    results_df = run_multiple_experiments(num_runs=20)

    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")

    save_results_to_excel(results_df, output_path)

    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)

    return results_df, output_path


if __name__ == "__main__":
    results, output_file = main()