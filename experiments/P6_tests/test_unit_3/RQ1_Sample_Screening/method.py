import os
import random
from collections import deque
import numpy as np
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# --- 全局状态范围配置 ---
STATE_MIN_X, STATE_MAX_X = 2000, 9000
STATE_MIN_Y, STATE_MAX_Y = 38, 85
STATE_MIN_Z, STATE_MAX_Z = 650, 1900


def generate_input():
    return [
        random.randint(STATE_MIN_X, STATE_MAX_X),
        random.randint(STATE_MIN_Y, STATE_MAX_Y),
        random.randint(STATE_MIN_Z, STATE_MAX_Z)
    ]


def execute_Tr(light, moisture, co2):
    actions = []
    triggered = set()

    # Fixed all if statements - using triggered.add() instead of b[0]=1
    if (light < 3500 and moisture > 75) != (light < 2500 and moisture > 75):
        triggered.add(1)
    if (light < 3500 and moisture > 75) != (light < 4500 and moisture > 75):
        triggered.add(2)
    if (light < 3500 and moisture > 75) != (light < 5500 and moisture > 75):
        triggered.add(3)
    if (light < 3500 and moisture > 75) != (light < 3500 and moisture > 25):
        triggered.add(4)
    if (light < 3500 and moisture > 75) != (light < 3500 and moisture > 55):
        triggered.add(5)

    if (light < 2500 and moisture > 80 and co2 > 1600) != (light < 5500 and moisture > 80 and co2 > 1600):
        triggered.add(6)
    if (light < 2500 and moisture > 80 and co2 > 1600) != (light < 6500 and moisture > 80 and co2 > 1600):
        triggered.add(7)
    if (light < 2500 and moisture > 80 and co2 > 1600) != (light < 8500 and moisture > 80 and co2 > 1600):
        triggered.add(8)
    if (light < 2500 and moisture > 80 and co2 > 1600) != (light < 2500 and moisture > 40 and co2 > 1600):
        triggered.add(9)
    # 原分支10已删除

    if (light > 3000 and moisture < 75) != (light > 3000 and moisture < 55):
        triggered.add(10)
    if (light > 3000 and moisture < 75) != (light > 4000 and moisture < 75):
        triggered.add(11)
    if (light > 3000 and moisture < 75) != (light > 5000 and moisture < 75):
        triggered.add(12)
    if (light > 3000 and moisture < 75) != (light > 6000 and moisture < 75):
        triggered.add(13)
    if (light > 3000 and moisture < 75) != (light > 3000 and moisture < 85):
        triggered.add(14)
    if (light > 3000 and moisture < 75) != (light > 3000 and moisture < 35):
        triggered.add(15)
    if (light > 3000 and moisture < 75) != (light > 3000 and moisture < 15):
        triggered.add(16)
    if (light > 3000 and moisture < 75) != (light > 3000 and moisture < 65):
        triggered.add(17)

    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 4500 or co2 < 800) and moisture > 42):
        triggered.add(18)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 6500 or co2 < 800) and moisture > 42):
        triggered.add(19)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 8500 or co2 < 800) and moisture > 42):
        triggered.add(20)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 3500 or co2 < 900) and moisture > 42):
        triggered.add(21)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 3500 or co2 < 1100) and moisture > 42):
        triggered.add(22)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 3500 or co2 < 700) and moisture > 42):
        triggered.add(23)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 3500 or co2 < 1500) and moisture > 42):
        triggered.add(24)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 3500 or co2 < 800) and moisture > 32):
        triggered.add(25)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 3500 or co2 < 800) and moisture > 52):
        triggered.add(26)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 3500 or co2 < 800) and moisture > 62):
        triggered.add(27)

    if (moisture / (light / 100) > 8) != (moisture / (co2 / 100) > 8):
        triggered.add(28)

    if (light < 3000 and (light + moisture * 10) < 3800) != (light < 2000 and (light + moisture * 10) < 3800):
        triggered.add(29)
    if (light < 3000 and (light + moisture * 10) < 3800) != (light < 4000 and (light + moisture * 10) < 3800):
        triggered.add(30)
    if (light < 3000 and (light + moisture * 10) < 3800) != (light < 5000 and (light + moisture * 10) < 3800):
        triggered.add(31)
    if (light < 3000 and (light + moisture * 10) < 3800) != (light < 7000 and (light + moisture * 10) < 3800):
        triggered.add(32)
    if (light < 3000 and (light + moisture * 10) < 3800) != (light < 3000 and (light + moisture * 15) < 3800):
        triggered.add(33)
    if (light < 3000 and (light + moisture * 10) < 3800) != (light < 3000 and (light + moisture * 20) < 3800):
        triggered.add(34)
    if (light < 3000 and (light + moisture * 10) < 3800) != (light < 3000 and (light + moisture * 10) > 3800):
        triggered.add(35)
    if (light < 3000 and (light + moisture * 10) < 3800) != (light > 3000 and (light + moisture * 10) < 3800):
        triggered.add(36)
    # 原分支38已删除
    if (light < 3000 and (light + moisture * 10) < 3800) != (light < 3000 and (light + co2) < 3800):
        triggered.add(37)

    if (light > 5000 and light < 7000 and co2 > 1400) != (light < 5000 and light < 7000 and co2 > 1400):
        triggered.add(38)
    if (light > 5000 and light < 7000 and co2 > 1400) != (light > 4000 and light < 7000 and co2 > 1400):
        triggered.add(39)
    if (light > 5000 and light < 7000 and co2 > 1400) != (light > 5000 and light > 7000 and co2 > 1400):
        triggered.add(40)
    if (light > 5000 and light < 7000 and co2 > 1400) != (light > 5000 and light < 3000 and co2 > 1400):
        triggered.add(41)
    if (light > 5000 and light < 7000 and co2 > 1400) != (light > 5000 and light < 7000 and co2 < 1400):
        triggered.add(42)
    if (light > 5000 and light < 7000 and co2 > 1400) != (light > 5000 and light < 7000 and co2 > 1200):
        triggered.add(43)
    if (light > 5000 and light < 7000 and co2 > 1400) != (light > 5000 and light < 4000 and co2 > 1400):
        triggered.add(44)

    if (light > 6000 and (light + moisture + co2) > 8500) != (light < 6000 and (light + moisture + co2) > 8500):
        triggered.add(45)
    if (light > 6000 and (light + moisture + co2) > 8500) != (light > 6000 and (light - moisture + co2) > 8500):
        triggered.add(46)
    if (light > 6000 and (light + moisture + co2) > 8500) != (light > 6000 and (light + moisture - co2) > 8500):
        triggered.add(47)
    if (light > 6000 and (light + moisture + co2) > 8500) != (light > 6000 and (light + moisture + co2) > 4500):
        triggered.add(48)
    if (light > 6000 and (light + moisture + co2) > 8500) != (light > 6000 and (light + moisture + co2) > 6500):
        triggered.add(49)
    if (light > 6000 and (light + moisture + co2) > 8500) != (light > 8000 and (light + moisture + co2) > 8500):
        triggered.add(50)

    if (light > 8500 and moisture < 50) != (light > 5500 and moisture < 50):
        triggered.add(51)
    if (light > 8500 and moisture < 50) != (light > 3500 and moisture < 50):
        triggered.add(52)
    if (light > 8500 and moisture < 50) != (light < 8500 and moisture < 50):
        triggered.add(53)
    if (light > 8500 and moisture < 50) != (light > 8500 and moisture < 70):
        triggered.add(54)
    if (light > 8500 and moisture < 50) != (light > 8500 and moisture < 60):
        triggered.add(55)
    if (light > 8500 and moisture < 50) != (light > 8500 and moisture > 50):
        triggered.add(56)
    if (light > 8500 and moisture < 50) != (light > 8500 and moisture < 40):
        triggered.add(57)
    if (light > 8500 and moisture < 50) != (light > 8500 and moisture < 30):
        triggered.add(58)

    if (light > 7500 and co2 < 900) != (light > 5500 and co2 < 900):
        triggered.add(59)
    if (light > 7500 and co2 < 900) != (light < 7500 and co2 < 900):
        triggered.add(60)
    if (light > 7500 and co2 < 900) != (light > 7500 and co2 > 900):
        triggered.add(61)
    if (light > 7500 and co2 < 900) != (light > 7500 and co2 < 400):
        triggered.add(62)
    if (light > 7500 and co2 < 900) != (light > 7000 and co2 < 900):
        triggered.add(63)
    if (light > 7500 and co2 < 900) != (light > 3500 and co2 < 900):
        triggered.add(64)
    if (light > 7500 and co2 < 900) != (light > 7500 and co2 < 1400):
        triggered.add(65)
    if (light > 7500 and co2 < 900) != (light > 7500 and co2 < 800):
        triggered.add(66)

    if (light < 4000 and moisture < 45) != (light < 3000 and moisture < 45):
        triggered.add(67)
    if (light < 4000 and moisture < 45) != (light < 5000 and moisture < 45):
        triggered.add(68)
    if (light < 4000 and moisture < 45) != (light < 6000 and moisture < 45):
        triggered.add(69)
    if (light < 4000 and moisture < 45) != (light > 4000 and moisture < 45):
        triggered.add(70)
    if (light < 4000 and moisture < 45) != (light < 4000 and moisture > 45):
        triggered.add(71)
    if (light < 4000 and moisture < 45) != (light < 4000 and moisture < 55):
        triggered.add(72)
    if (light < 4000 and moisture < 45) != (light < 4000 and moisture < 65):
        triggered.add(73)
    if (light < 4000 and moisture < 45) != (light < 4000 and moisture < 75):
        triggered.add(74)
    if (light < 4000 and moisture < 45) != (light < 7000 and moisture < 45):
        triggered.add(75)

    if (light < 2800 and moisture > 80) != (light < 1800 and moisture > 80):
        triggered.add(76)
    if (light < 2800 and moisture > 80) != (light > 2800 and moisture > 80):
        triggered.add(77)
    if (light < 2800 and moisture > 80) != (light < 3800 and moisture > 80):
        triggered.add(78)
    if (light < 2800 and moisture > 80) != (light < 4800 and moisture > 80):
        triggered.add(79)
    if (light < 2800 and moisture > 80) != (light < 2800 and moisture < 80):
        triggered.add(80)
    if (light < 2800 and moisture > 80) != (light < 2800 and moisture > 60):
        triggered.add(81)
    if (light < 2800 and moisture > 80) != (light < 2800 and moisture > 40):
        triggered.add(82)
    if (light < 2800 and moisture > 80) != (light < 2800 and moisture > 30):
        triggered.add(83)

    if (light / (co2 + 1) > 7) != (light / (co2 + 100) > 7):
        triggered.add(84)
    if (light / (co2 + 1) > 7) != (light / (co2 + 200) > 7):
        triggered.add(85)
    if (light / (co2 + 1) > 7) != (light / (co2 + 1) > 8):
        triggered.add(86)
    if (light / (co2 + 1) > 7) != (light / (co2 - 1) > 7):
        triggered.add(87)
    if (light / (co2 + 1) > 7) != (light / (co2 + 1) > 9):
        triggered.add(88)
    if (light / (co2 + 1) > 7) != (light / (co2 + 300) > 7):
        triggered.add(89)
    if (light / (co2 + 1) > 7) != (light / (co2 - 100) > 7):
        triggered.add(90)
    if (light / (co2 + 1) > 7) != (light / (co2 - 200) > 7):
        triggered.add(91)
    if (light / (co2 + 1) > 7) != (light / (co2 - 300) > 7):
        triggered.add(92)
    if (light / (co2 + 1) > 7) != (light / (co2 + 1) > 10):
        triggered.add(93)

    if (moisture < 55 and light < 4500) != (moisture < 65 and light < 4500):
        triggered.add(94)
    if (moisture < 55 and light < 4500) != (moisture < 75 and light < 4500):
        triggered.add(95)
    if (moisture < 55 and light < 4500) != (moisture < 85 and light < 4500):
        triggered.add(96)
    if (moisture < 55 and light < 4500) != (moisture > 55 and light < 4500):
        triggered.add(97)
    if (moisture < 55 and light < 4500) != (moisture < 55 and light > 4500):
        triggered.add(98)
    if (moisture < 55 and light < 4500) != (moisture < 55 and light < 3500):
        triggered.add(99)
    if (moisture < 55 and light < 4500) != (moisture < 55 and light < 6500):
        triggered.add(100)
    if (moisture < 55 and light < 4500) != (moisture < 55 and light < 7500):
        triggered.add(101)
    if (moisture < 55 and light < 4500) != (moisture < 55 and light < 8500):
        triggered.add(102)

    if (moisture < 58 and co2 < 950) != (moisture < 38 and co2 < 950):
        triggered.add(103)
    if (moisture < 58 and co2 < 950) != (moisture < 48 and co2 < 950):
        triggered.add(104)
    if (moisture < 58 and co2 < 950) != (moisture > 58 and co2 < 950):
        triggered.add(105)
    if (moisture < 58 and co2 < 950) != (moisture < 58 and co2 > 950):
        triggered.add(106)
    if (moisture < 58 and co2 < 950) != (moisture < 58 and co2 < 1050):
        triggered.add(107)
    if (moisture < 58 and co2 < 950) != (moisture < 58 and co2 < 450):
        triggered.add(108)

    if (moisture > 82 and co2 > 1600) != (moisture > 62 and co2 > 1600):
        triggered.add(109)
    if (moisture > 82 and co2 > 1600) != (moisture > 32 and co2 > 1600):
        triggered.add(110)
    if (moisture > 82 and co2 > 1600) != (moisture > 42 and co2 > 1600):
        triggered.add(111)
    if (moisture > 82 and co2 > 1600) != (moisture > 82 and co2 > 1100):
        triggered.add(112)
    if (moisture > 82 and co2 > 1600) != (moisture > 82 and co2 > 1000):
        triggered.add(113)
    if (moisture > 82 and co2 > 1600) != (moisture > 82 and co2 > 700):
        triggered.add(114)

    if (co2 > 1750 and moisture < 45) != (co2 > 950 and moisture < 45):
        triggered.add(115)
    if (co2 > 1750 and moisture < 45) != (co2 > 750 and moisture < 45):
        triggered.add(116)
    if (co2 > 1750 and moisture < 45) != (co2 < 1750 and moisture < 45):
        triggered.add(117)
    if (co2 > 1750 and moisture < 45) != (co2 > 1750 and moisture > 45):
        triggered.add(118)
    if (co2 > 1750 and moisture < 45) != (co2 > 1750 and moisture < 55):
        triggered.add(119)
    if (co2 > 1750 and moisture < 45) != (co2 > 1750 and moisture < 65):
        triggered.add(120)

    if (co2 / (light + 1) > 0.22) != (co2 / (light + 100) > 0.22):
        triggered.add(121)
    if (co2 / (light + 1) > 0.22) != (co2 / (light + 200) > 0.22):
        triggered.add(122)
    if (co2 / (light + 1) > 0.22) != (co2 / (light - 1000) > 0.22):
        triggered.add(123)
    if (co2 / (light + 1) > 0.22) != (co2 / (light - 2000) > 0.22):
        triggered.add(124)
    if (co2 / (light + 1) > 0.22) != (co2 / (light + 1) > 0.32):
        triggered.add(125)

    if (light + moisture * 25 + co2 > 10500) != (light + moisture * 15 + co2 > 10500):
        triggered.add(126)
    if (light + moisture * 25 + co2 > 10500) != (light + moisture * 35 + co2 > 10500):
        triggered.add(127)
    if (light + moisture * 25 + co2 > 10500) != (light + moisture * 45 + co2 > 10500):
        triggered.add(128)
    if (light + moisture * 25 + co2 > 10500) != (light + moisture * 25 + co2 + 1000 > 10500):
        triggered.add(129)

    return triggered


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0


# Fixed: Convert all sets to lists for consistency
TARGET_PATHS = [
    [15, 16, 19, 20, 21, 22, 24, 42, 48, 49, 51, 52, 53, 59, 60, 64, 70, 75, 84, 85, 86, 88, 89, 93, 98, 100, 101, 102,
     103, 105, 106, 108, 116, 117],
    [15, 16, 23, 26, 27, 42, 48, 49, 51, 52, 53, 59, 60, 64, 70, 75, 84, 85, 86, 88, 89, 93, 98, 100, 101, 102, 103,
     105, 106, 108, 116, 117],
    [15, 16, 19, 20, 21, 22, 24, 42, 48, 49, 51, 52, 53, 59, 60, 64, 70, 75, 90, 91, 92, 98, 100, 101, 102, 103, 105,
     106, 108, 116, 117],
    [13, 15, 16, 19, 20, 21, 22, 24, 42, 52, 53, 60, 64, 69, 70, 75, 91, 92, 98, 100, 101, 102, 103, 105, 106, 108, 116,
     117, 123, 124],
    [12, 13, 15, 16, 19, 20, 21, 22, 24, 52, 53, 60, 64, 68, 69, 70, 75, 91, 92, 98, 100, 101, 102, 103, 105, 106, 108,
     116, 117, 124],
    [11, 12, 13, 15, 16, 18, 19, 20, 21, 22, 24, 52, 53, 60, 64, 67, 70, 71, 92, 97, 98, 99, 103, 105, 106, 108, 116,
     117, 123, 124],
    [15, 16, 20, 22, 24, 45, 46, 47, 50, 51, 52, 53, 61, 65, 70, 84, 85, 86, 88, 89, 93, 98, 102, 106, 107, 115, 116,
     117, 128, 129],
    [11, 12, 13, 15, 16, 18, 19, 20, 21, 22, 24, 52, 53, 60, 64, 71, 72, 73, 74, 92, 97, 98, 99, 103, 104, 105, 106,
     108, 123, 124],
    [11, 12, 13, 15, 16, 18, 19, 20, 21, 22, 24, 60, 64, 71, 72, 73, 74, 92, 97, 98, 99, 103, 104, 105, 106, 108, 121,
     122, 125],
    [15, 16, 20, 22, 24, 45, 47, 50, 51, 52, 53, 61, 65, 70, 84, 85, 86, 88, 89, 93, 98, 102, 106, 115, 116, 117, 127,
     128, 129],
    [15, 16, 20, 38, 40, 41, 42, 44, 45, 47, 50, 51, 52, 53, 70, 75, 98, 101, 102, 106, 110, 111, 115, 116, 117, 125,
     128, 129],
    [4, 11, 12, 13, 15, 16, 26, 27, 30, 31, 32, 36, 53, 60, 71, 72, 73, 74, 92, 97, 98, 103, 104, 105, 106, 108, 121,
     122, 125],
    [15, 16, 20, 21, 22, 24, 45, 47, 51, 52, 53, 60, 61, 62, 66, 70, 93, 98, 102, 103, 105, 106, 108, 116, 117, 127,
     128, 129],
    [15, 16, 20, 38, 40, 41, 42, 44, 45, 46, 47, 50, 51, 52, 53, 70, 75, 98, 101, 102, 106, 110, 111, 117, 118, 125,
     128, 129],
    [13, 15, 16, 25, 42, 51, 52, 53, 59, 60, 64, 69, 70, 75, 85, 86, 88, 89, 93, 98, 100, 101, 102, 103, 105, 106, 108,
     117],
    [15, 16, 20, 21, 22, 24, 48, 49, 51, 52, 53, 59, 60, 63, 64, 85, 88, 89, 93, 98, 101, 102, 103, 104, 105, 106, 108,
     129],
    [15, 16, 20, 38, 40, 41, 42, 44, 45, 46, 47, 50, 51, 52, 53, 98, 101, 102, 106, 110, 111, 118, 119, 120, 125, 128,
     129],
    [13, 15, 16, 42, 51, 52, 53, 59, 60, 64, 69, 70, 75, 87, 90, 91, 92, 98, 100, 101, 102, 103, 105, 106, 108, 116,
     117],
    [15, 16, 19, 20, 24, 42, 43, 48, 49, 51, 52, 53, 70, 75, 92, 98, 100, 101, 102, 106, 115, 116, 117, 123, 124],
    [4, 5, 10, 11, 12, 13, 15, 16, 27, 28, 60, 71, 73, 74, 92, 94, 95, 96, 97, 103, 104, 105, 106, 108, 123, 124],
    [15, 16, 23, 26, 27, 45, 47, 53, 56, 57, 58, 60, 61, 62, 70, 98, 103, 105, 106, 108, 116, 117, 127, 128, 129],
    [4, 5, 27, 28, 29, 34, 35, 36, 60, 71, 73, 74, 80, 82, 83, 92, 94, 95, 96, 97, 103, 104, 105, 106, 108, 125],
    [12, 13, 15, 16, 19, 20, 24, 38, 39, 52, 53, 68, 69, 70, 75, 98, 100, 101, 102, 106, 115, 116, 117, 125],
    [10, 15, 16, 17, 20, 21, 22, 24, 28, 45, 46, 47, 50, 60, 61, 62, 66, 85, 88, 89, 93, 105, 127, 128, 129],
    [15, 16, 21, 22, 24, 45, 47, 53, 56, 57, 58, 60, 61, 62, 66, 70, 98, 103, 105, 106, 108, 116, 117, 126],
    [4, 5, 10, 11, 12, 13, 15, 16, 30, 31, 32, 36, 38, 71, 73, 74, 94, 95, 96, 97, 109, 110, 111, 118, 120],
    [4, 9, 26, 27, 29, 35, 36, 37, 38, 53, 70, 71, 80, 82, 83, 97, 98, 106, 110, 111, 115, 116, 117],
    [15, 16, 21, 22, 24, 45, 47, 54, 55, 56, 60, 61, 62, 66, 93, 98, 103, 104, 105, 106, 108, 126],
    [14, 19, 20, 24, 38, 40, 41, 42, 44, 48, 49, 77, 112, 113, 114, 121, 122, 125, 127, 128, 129],
    [4, 5, 27, 28, 29, 33, 34, 35, 36, 60, 71, 73, 74, 92, 94, 95, 96, 97, 105, 121, 122, 125],
    [4, 5, 29, 33, 34, 35, 36, 37, 38, 71, 74, 80, 81, 82, 83, 95, 96, 97, 109, 110, 111, 118],
    [2, 3, 14, 18, 19, 20, 21, 22, 24, 28, 60, 64, 77, 79, 96, 97, 105, 114, 121, 122, 125],
    [2, 3, 14, 18, 19, 20, 21, 22, 24, 28, 60, 64, 71, 77, 78, 79, 96, 97, 105, 122, 125],
    [1, 6, 7, 8, 29, 33, 34, 35, 36, 37, 38, 71, 77, 78, 79, 96, 97, 109, 110, 111, 118],
    [1, 28, 29, 33, 34, 35, 36, 60, 71, 76, 77, 80, 96, 97, 105, 114, 125]
]


class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 2000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(TARGET_PATHS)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }


def compute_robustness(state, path):
    base = execute_Tr(state[0], state[1], state[2])
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                    np.clip(state[0] + dx, STATE_MIN_X, STATE_MAX_X),
                    np.clip(state[1] + dy, STATE_MIN_Y, STATE_MAX_Y),
                    np.clip(state[2] + dz, STATE_MIN_Z, STATE_MAX_Z)
                ])
                n_trig = execute_Tr(neighbor[0], neighbor[1], neighbor[2])
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = TARGET_PATHS[target_path_idx]
    samples = []
    attempts = 0

    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(STATE_MIN_X, STATE_MAX_X),
            random.randint(STATE_MIN_Y, STATE_MAX_Y),
            random.randint(STATE_MIN_Z, STATE_MAX_Z)
        ])
        triggered = execute_Tr(state[0], state[1], state[2])

        if not triggered:
            continue

        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)

        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }

        samples.append(sample_data)

    return samples


def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample[
                'robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)

        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]

    return selected_samples


def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = TARGET_PATHS[target_path_idx]
        samples = []
        attempts = 0

        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(STATE_MIN_X, STATE_MAX_X),
                random.randint(STATE_MIN_Y, STATE_MAX_Y),
                random.randint(STATE_MIN_Z, STATE_MAX_Z)
            ])
            triggered = execute_Tr(state[0], state[1], state[2])

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)

            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)

        random.shuffle(samples)
        return samples[:config.top_k_samples]

    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")

        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)


def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}

    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates

    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}

        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config,
                                                         shared_candidates[path_idx])

            strategy_results[path_idx] = samples

        results[strategy_name] = strategy_results

    return results


def analyze_fitness_values(results, config):
    analysis_results = {}

    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []

        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])

        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }

        all_scores = []

        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])

        scores_array = np.array(all_scores)

        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)

        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })

        analysis_results[strategy_name] = analysis

    return analysis_results


def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []

    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)

    df = pd.DataFrame(df_data)
    return df, analysis_results


def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []

    print(f"Starting {num_runs} experiments...")

    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")

        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)

        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)

        print(f"Experiment {run_idx + 1} completed.")

    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df


def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')

    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width

    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []

    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]

        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)

    stats_df = pd.DataFrame(stats_data)

    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Results saved to: {output_path}")


def main():
    results_df = run_multiple_experiments(num_runs=20)

    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")

    save_results_to_excel(results_df, output_path)

    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)

    return results_df, output_path


if __name__ == "__main__":
    results, output_file = main()