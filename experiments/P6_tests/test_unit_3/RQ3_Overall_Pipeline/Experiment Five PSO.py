import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

def execute_validation_rules(light, moisture, co2):
    actions = []
    triggered = set()

    # Fixed all if statements - using triggered.add() instead of b[0]=1
    if (light < 3500 and moisture > 75) != (light < 2500 and moisture > 75):
        triggered.add(1)
    if (light < 3500 and moisture > 75) != (light < 4500 and moisture > 75):
        triggered.add(2)
    if (light < 3500 and moisture > 75) != (light < 5500 and moisture > 75):
        triggered.add(3)
    if (light < 3500 and moisture > 75) != (light < 3500 and moisture > 25):
        triggered.add(4)
    if (light < 3500 and moisture > 75) != (light < 3500 and moisture > 55):
        triggered.add(5)

    if (light < 2500 and moisture > 80 and co2 > 1600) != (light < 5500 and moisture > 80 and co2 > 1600):
        triggered.add(6)
    if (light < 2500 and moisture > 80 and co2 > 1600) != (light < 6500 and moisture > 80 and co2 > 1600):
        triggered.add(7)
    if (light < 2500 and moisture > 80 and co2 > 1600) != (light < 8500 and moisture > 80 and co2 > 1600):
        triggered.add(8)
    if (light < 2500 and moisture > 80 and co2 > 1600) != (light < 2500 and moisture > 40 and co2 > 1600):
        triggered.add(9)
    # 原分支10已删除

    if (light > 3000 and moisture < 75) != (light > 3000 and moisture < 55):
        triggered.add(10)
    if (light > 3000 and moisture < 75) != (light > 4000 and moisture < 75):
        triggered.add(11)
    if (light > 3000 and moisture < 75) != (light > 5000 and moisture < 75):
        triggered.add(12)
    if (light > 3000 and moisture < 75) != (light > 6000 and moisture < 75):
        triggered.add(13)
    if (light > 3000 and moisture < 75) != (light > 3000 and moisture < 85):
        triggered.add(14)
    if (light > 3000 and moisture < 75) != (light > 3000 and moisture < 35):
        triggered.add(15)
    if (light > 3000 and moisture < 75) != (light > 3000 and moisture < 15):
        triggered.add(16)
    if (light > 3000 and moisture < 75) != (light > 3000 and moisture < 65):
        triggered.add(17)

    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 4500 or co2 < 800) and moisture > 42):
        triggered.add(18)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 6500 or co2 < 800) and moisture > 42):
        triggered.add(19)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 8500 or co2 < 800) and moisture > 42):
        triggered.add(20)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 3500 or co2 < 900) and moisture > 42):
        triggered.add(21)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 3500 or co2 < 1100) and moisture > 42):
        triggered.add(22)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 3500 or co2 < 700) and moisture > 42):
        triggered.add(23)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 3500 or co2 < 1500) and moisture > 42):
        triggered.add(24)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 3500 or co2 < 800) and moisture > 32):
        triggered.add(25)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 3500 or co2 < 800) and moisture > 52):
        triggered.add(26)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 3500 or co2 < 800) and moisture > 62):
        triggered.add(27)

    if (moisture / (light / 100) > 8) != (moisture / (co2 / 100) > 8):
        triggered.add(28)

    if (light < 3000 and (light + moisture * 10) < 3800) != (light < 2000 and (light + moisture * 10) < 3800):
        triggered.add(29)
    if (light < 3000 and (light + moisture * 10) < 3800) != (light < 4000 and (light + moisture * 10) < 3800):
        triggered.add(30)
    if (light < 3000 and (light + moisture * 10) < 3800) != (light < 5000 and (light + moisture * 10) < 3800):
        triggered.add(31)
    if (light < 3000 and (light + moisture * 10) < 3800) != (light < 7000 and (light + moisture * 10) < 3800):
        triggered.add(32)
    if (light < 3000 and (light + moisture * 10) < 3800) != (light < 3000 and (light + moisture * 15) < 3800):
        triggered.add(33)
    if (light < 3000 and (light + moisture * 10) < 3800) != (light < 3000 and (light + moisture * 20) < 3800):
        triggered.add(34)
    if (light < 3000 and (light + moisture * 10) < 3800) != (light < 3000 and (light + moisture * 10) > 3800):
        triggered.add(35)
    if (light < 3000 and (light + moisture * 10) < 3800) != (light > 3000 and (light + moisture * 10) < 3800):
        triggered.add(36)
    # 原分支38已删除
    if (light < 3000 and (light + moisture * 10) < 3800) != (light < 3000 and (light + co2) < 3800):
        triggered.add(37)

    if (light > 5000 and light < 7000 and co2 > 1400) != (light < 5000 and light < 7000 and co2 > 1400):
        triggered.add(38)
    if (light > 5000 and light < 7000 and co2 > 1400) != (light > 4000 and light < 7000 and co2 > 1400):
        triggered.add(39)
    if (light > 5000 and light < 7000 and co2 > 1400) != (light > 5000 and light > 7000 and co2 > 1400):
        triggered.add(40)
    if (light > 5000 and light < 7000 and co2 > 1400) != (light > 5000 and light < 3000 and co2 > 1400):
        triggered.add(41)
    if (light > 5000 and light < 7000 and co2 > 1400) != (light > 5000 and light < 7000 and co2 < 1400):
        triggered.add(42)
    if (light > 5000 and light < 7000 and co2 > 1400) != (light > 5000 and light < 7000 and co2 > 1200):
        triggered.add(43)
    if (light > 5000 and light < 7000 and co2 > 1400) != (light > 5000 and light < 4000 and co2 > 1400):
        triggered.add(44)

    if (light > 6000 and (light + moisture + co2) > 8500) != (light < 6000 and (light + moisture + co2) > 8500):
        triggered.add(45)
    if (light > 6000 and (light + moisture + co2) > 8500) != (light > 6000 and (light - moisture + co2) > 8500):
        triggered.add(46)
    if (light > 6000 and (light + moisture + co2) > 8500) != (light > 6000 and (light + moisture - co2) > 8500):
        triggered.add(47)
    if (light > 6000 and (light + moisture + co2) > 8500) != (light > 6000 and (light + moisture + co2) > 4500):
        triggered.add(48)
    if (light > 6000 and (light + moisture + co2) > 8500) != (light > 6000 and (light + moisture + co2) > 6500):
        triggered.add(49)
    if (light > 6000 and (light + moisture + co2) > 8500) != (light > 8000 and (light + moisture + co2) > 8500):
        triggered.add(50)

    if (light > 8500 and moisture < 50) != (light > 5500 and moisture < 50):
        triggered.add(51)
    if (light > 8500 and moisture < 50) != (light > 3500 and moisture < 50):
        triggered.add(52)
    if (light > 8500 and moisture < 50) != (light < 8500 and moisture < 50):
        triggered.add(53)
    if (light > 8500 and moisture < 50) != (light > 8500 and moisture < 70):
        triggered.add(54)
    if (light > 8500 and moisture < 50) != (light > 8500 and moisture < 60):
        triggered.add(55)
    if (light > 8500 and moisture < 50) != (light > 8500 and moisture > 50):
        triggered.add(56)
    if (light > 8500 and moisture < 50) != (light > 8500 and moisture < 40):
        triggered.add(57)
    if (light > 8500 and moisture < 50) != (light > 8500 and moisture < 30):
        triggered.add(58)

    if (light > 7500 and co2 < 900) != (light > 5500 and co2 < 900):
        triggered.add(59)
    if (light > 7500 and co2 < 900) != (light < 7500 and co2 < 900):
        triggered.add(60)
    if (light > 7500 and co2 < 900) != (light > 7500 and co2 > 900):
        triggered.add(61)
    if (light > 7500 and co2 < 900) != (light > 7500 and co2 < 400):
        triggered.add(62)
    if (light > 7500 and co2 < 900) != (light > 7000 and co2 < 900):
        triggered.add(63)
    if (light > 7500 and co2 < 900) != (light > 3500 and co2 < 900):
        triggered.add(64)
    if (light > 7500 and co2 < 900) != (light > 7500 and co2 < 1400):
        triggered.add(65)
    if (light > 7500 and co2 < 900) != (light > 7500 and co2 < 800):
        triggered.add(66)

    if (light < 4000 and moisture < 45) != (light < 3000 and moisture < 45):
        triggered.add(67)
    if (light < 4000 and moisture < 45) != (light < 5000 and moisture < 45):
        triggered.add(68)
    if (light < 4000 and moisture < 45) != (light < 6000 and moisture < 45):
        triggered.add(69)
    if (light < 4000 and moisture < 45) != (light > 4000 and moisture < 45):
        triggered.add(70)
    if (light < 4000 and moisture < 45) != (light < 4000 and moisture > 45):
        triggered.add(71)
    if (light < 4000 and moisture < 45) != (light < 4000 and moisture < 55):
        triggered.add(72)
    if (light < 4000 and moisture < 45) != (light < 4000 and moisture < 65):
        triggered.add(73)
    if (light < 4000 and moisture < 45) != (light < 4000 and moisture < 75):
        triggered.add(74)
    if (light < 4000 and moisture < 45) != (light < 7000 and moisture < 45):
        triggered.add(75)

    if (light < 2800 and moisture > 80) != (light < 1800 and moisture > 80):
        triggered.add(76)
    if (light < 2800 and moisture > 80) != (light > 2800 and moisture > 80):
        triggered.add(77)
    if (light < 2800 and moisture > 80) != (light < 3800 and moisture > 80):
        triggered.add(78)
    if (light < 2800 and moisture > 80) != (light < 4800 and moisture > 80):
        triggered.add(79)
    if (light < 2800 and moisture > 80) != (light < 2800 and moisture < 80):
        triggered.add(80)
    if (light < 2800 and moisture > 80) != (light < 2800 and moisture > 60):
        triggered.add(81)
    if (light < 2800 and moisture > 80) != (light < 2800 and moisture > 40):
        triggered.add(82)
    if (light < 2800 and moisture > 80) != (light < 2800 and moisture > 30):
        triggered.add(83)

    if (light / (co2 + 1) > 7) != (light / (co2 + 100) > 7):
        triggered.add(84)
    if (light / (co2 + 1) > 7) != (light / (co2 + 200) > 7):
        triggered.add(85)
    if (light / (co2 + 1) > 7) != (light / (co2 + 1) > 8):
        triggered.add(86)
    if (light / (co2 + 1) > 7) != (light / (co2 - 1) > 7):
        triggered.add(87)
    if (light / (co2 + 1) > 7) != (light / (co2 + 1) > 9):
        triggered.add(88)
    if (light / (co2 + 1) > 7) != (light / (co2 + 300) > 7):
        triggered.add(89)
    if (light / (co2 + 1) > 7) != (light / (co2 - 100) > 7):
        triggered.add(90)
    if (light / (co2 + 1) > 7) != (light / (co2 - 200) > 7):
        triggered.add(91)
    if (light / (co2 + 1) > 7) != (light / (co2 - 300) > 7):
        triggered.add(92)
    if (light / (co2 + 1) > 7) != (light / (co2 + 1) > 10):
        triggered.add(93)

    if (moisture < 55 and light < 4500) != (moisture < 65 and light < 4500):
        triggered.add(94)
    if (moisture < 55 and light < 4500) != (moisture < 75 and light < 4500):
        triggered.add(95)
    if (moisture < 55 and light < 4500) != (moisture < 85 and light < 4500):
        triggered.add(96)
    if (moisture < 55 and light < 4500) != (moisture > 55 and light < 4500):
        triggered.add(97)
    if (moisture < 55 and light < 4500) != (moisture < 55 and light > 4500):
        triggered.add(98)
    if (moisture < 55 and light < 4500) != (moisture < 55 and light < 3500):
        triggered.add(99)
    if (moisture < 55 and light < 4500) != (moisture < 55 and light < 6500):
        triggered.add(100)
    if (moisture < 55 and light < 4500) != (moisture < 55 and light < 7500):
        triggered.add(101)
    if (moisture < 55 and light < 4500) != (moisture < 55 and light < 8500):
        triggered.add(102)

    if (moisture < 58 and co2 < 950) != (moisture < 38 and co2 < 950):
        triggered.add(103)
    if (moisture < 58 and co2 < 950) != (moisture < 48 and co2 < 950):
        triggered.add(104)
    if (moisture < 58 and co2 < 950) != (moisture > 58 and co2 < 950):
        triggered.add(105)
    if (moisture < 58 and co2 < 950) != (moisture < 58 and co2 > 950):
        triggered.add(106)
    if (moisture < 58 and co2 < 950) != (moisture < 58 and co2 < 1050):
        triggered.add(107)
    if (moisture < 58 and co2 < 950) != (moisture < 58 and co2 < 450):
        triggered.add(108)

    if (moisture > 82 and co2 > 1600) != (moisture > 62 and co2 > 1600):
        triggered.add(109)
    if (moisture > 82 and co2 > 1600) != (moisture > 32 and co2 > 1600):
        triggered.add(110)
    if (moisture > 82 and co2 > 1600) != (moisture > 42 and co2 > 1600):
        triggered.add(111)
    if (moisture > 82 and co2 > 1600) != (moisture > 82 and co2 > 1100):
        triggered.add(112)
    if (moisture > 82 and co2 > 1600) != (moisture > 82 and co2 > 1000):
        triggered.add(113)
    if (moisture > 82 and co2 > 1600) != (moisture > 82 and co2 > 700):
        triggered.add(114)

    if (co2 > 1750 and moisture < 45) != (co2 > 950 and moisture < 45):
        triggered.add(115)
    if (co2 > 1750 and moisture < 45) != (co2 > 750 and moisture < 45):
        triggered.add(116)
    if (co2 > 1750 and moisture < 45) != (co2 < 1750 and moisture < 45):
        triggered.add(117)
    if (co2 > 1750 and moisture < 45) != (co2 > 1750 and moisture > 45):
        triggered.add(118)
    if (co2 > 1750 and moisture < 45) != (co2 > 1750 and moisture < 55):
        triggered.add(119)
    if (co2 > 1750 and moisture < 45) != (co2 > 1750 and moisture < 65):
        triggered.add(120)

    if (co2 / (light + 1) > 0.22) != (co2 / (light + 100) > 0.22):
        triggered.add(121)
    if (co2 / (light + 1) > 0.22) != (co2 / (light + 200) > 0.22):
        triggered.add(122)
    if (co2 / (light + 1) > 0.22) != (co2 / (light - 1000) > 0.22):
        triggered.add(123)
    if (co2 / (light + 1) > 0.22) != (co2 / (light - 2000) > 0.22):
        triggered.add(124)
    if (co2 / (light + 1) > 0.22) != (co2 / (light + 1) > 0.32):
        triggered.add(125)

    if (light + moisture * 25 + co2 > 10500) != (light + moisture * 15 + co2 > 10500):
        triggered.add(126)
    if (light + moisture * 25 + co2 > 10500) != (light + moisture * 35 + co2 > 10500):
        triggered.add(127)
    if (light + moisture * 25 + co2 > 10500) != (light + moisture * 45 + co2 > 10500):
        triggered.add(128)
    if (light + moisture * 25 + co2 > 10500) != (light + moisture * 25 + co2 + 1000 > 10500):
        triggered.add(129)

    return triggered

def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """"""
    generated_path = execute_validation_rules(particle[0], particle[1], particle[2])

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # : x:1-50, y:1-50, z:1-50
        self.bounds = bounds if bounds else [(1, 100), (1, 100), (1, 100)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """target pathsPSO"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(particles[i][0], particles[i][1], particles[i][2]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(gbest_particle[0], gbest_particle[1], gbest_particle[2]),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """Path PSO"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - Path ")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations")
    print(f"Path : {len(target_paths)}")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"Path {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "" if result['success'] else f"({result['best_fitness']:.3f})"
        print(f"{status} | {result['time']:.2f}s | iterations{result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f": {success_count}/{len(target_paths)} ({success_rate:.1f}%) | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs}")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations, {len(target_paths)}Path ")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"---  {run_idx}/{num_runs} ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f": {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs} runcompleted | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 1: 
    ws1 = wb.active
    ws1.title = ""
    ws1.sheet_view.showGridLines = False

    headers = ["Run", "", "", "", "Average Iterations", "(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f" {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # 2: Path 
    ws2 = wb.create_sheet(title="Path ")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Path ID", "", "", "", "Average Iterations", "Minimum Iterations", "Maximum Iterations"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"Path  {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # 3: 
    ws3 = wb.create_sheet(title="")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Path ", "", "(x,y,z)", "", "Iterations", "Path "]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"{run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # 4: target paths
    ws4 = wb.create_sheet(title="target paths")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Path ID", "target paths", ""]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"Path  {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f" : {filename}")
    print(f"{'=' * 70}")
    print(f":")
    print(f"  1.        - {len(all_results)} run")
    print(f"  2. Path        - Path ")
    print(f"  3.    -  runPath ")
    print(f"  4. target paths       - target paths")
    print(f"{'=' * 70}\n")

    return filename


def main():


    target_paths = [
        [15, 16, 19, 20, 21, 22, 24, 42, 48, 49, 51, 52, 53, 59, 60, 64, 70, 75, 84, 85, 86, 88, 89, 93, 98, 100, 101,
         102,
         103, 105, 106, 108, 116, 117],
        [15, 16, 23, 26, 27, 42, 48, 49, 51, 52, 53, 59, 60, 64, 70, 75, 84, 85, 86, 88, 89, 93, 98, 100, 101, 102, 103,
         105, 106, 108, 116, 117],
        [15, 16, 19, 20, 21, 22, 24, 42, 48, 49, 51, 52, 53, 59, 60, 64, 70, 75, 90, 91, 92, 98, 100, 101, 102, 103,
         105,
         106, 108, 116, 117],
        [13, 15, 16, 19, 20, 21, 22, 24, 42, 52, 53, 60, 64, 69, 70, 75, 91, 92, 98, 100, 101, 102, 103, 105, 106, 108,
         116,
         117, 123, 124],
        [12, 13, 15, 16, 19, 20, 21, 22, 24, 52, 53, 60, 64, 68, 69, 70, 75, 91, 92, 98, 100, 101, 102, 103, 105, 106,
         108,
         116, 117, 124],
        [11, 12, 13, 15, 16, 18, 19, 20, 21, 22, 24, 52, 53, 60, 64, 67, 70, 71, 92, 97, 98, 99, 103, 105, 106, 108,
         116,
         117, 123, 124],
        [15, 16, 20, 22, 24, 45, 46, 47, 50, 51, 52, 53, 61, 65, 70, 84, 85, 86, 88, 89, 93, 98, 102, 106, 107, 115,
         116,
         117, 128, 129],
        [11, 12, 13, 15, 16, 18, 19, 20, 21, 22, 24, 52, 53, 60, 64, 71, 72, 73, 74, 92, 97, 98, 99, 103, 104, 105, 106,
         108, 123, 124],
        [11, 12, 13, 15, 16, 18, 19, 20, 21, 22, 24, 60, 64, 71, 72, 73, 74, 92, 97, 98, 99, 103, 104, 105, 106, 108,
         121,
         122, 125],
        [15, 16, 20, 22, 24, 45, 47, 50, 51, 52, 53, 61, 65, 70, 84, 85, 86, 88, 89, 93, 98, 102, 106, 115, 116, 117,
         127,
         128, 129],
        [15, 16, 20, 38, 40, 41, 42, 44, 45, 47, 50, 51, 52, 53, 70, 75, 98, 101, 102, 106, 110, 111, 115, 116, 117,
         125,
         128, 129],
        [4, 11, 12, 13, 15, 16, 26, 27, 30, 31, 32, 36, 53, 60, 71, 72, 73, 74, 92, 97, 98, 103, 104, 105, 106, 108,
         121,
         122, 125],
        [15, 16, 20, 21, 22, 24, 45, 47, 51, 52, 53, 60, 61, 62, 66, 70, 93, 98, 102, 103, 105, 106, 108, 116, 117, 127,
         128, 129],
        [15, 16, 20, 38, 40, 41, 42, 44, 45, 46, 47, 50, 51, 52, 53, 70, 75, 98, 101, 102, 106, 110, 111, 117, 118, 125,
         128, 129],
        [13, 15, 16, 25, 42, 51, 52, 53, 59, 60, 64, 69, 70, 75, 85, 86, 88, 89, 93, 98, 100, 101, 102, 103, 105, 106,
         108,
         117],
        [15, 16, 20, 21, 22, 24, 48, 49, 51, 52, 53, 59, 60, 63, 64, 85, 88, 89, 93, 98, 101, 102, 103, 104, 105, 106,
         108,
         129],
        [15, 16, 20, 38, 40, 41, 42, 44, 45, 46, 47, 50, 51, 52, 53, 98, 101, 102, 106, 110, 111, 118, 119, 120, 125,
         128,
         129],
        [13, 15, 16, 42, 51, 52, 53, 59, 60, 64, 69, 70, 75, 87, 90, 91, 92, 98, 100, 101, 102, 103, 105, 106, 108, 116,
         117],
        [15, 16, 19, 20, 24, 42, 43, 48, 49, 51, 52, 53, 70, 75, 92, 98, 100, 101, 102, 106, 115, 116, 117, 123, 124],
        [4, 5, 10, 11, 12, 13, 15, 16, 27, 28, 60, 71, 73, 74, 92, 94, 95, 96, 97, 103, 104, 105, 106, 108, 123, 124],
        [15, 16, 23, 26, 27, 45, 47, 53, 56, 57, 58, 60, 61, 62, 70, 98, 103, 105, 106, 108, 116, 117, 127, 128, 129],
        [4, 5, 27, 28, 29, 34, 35, 36, 60, 71, 73, 74, 80, 82, 83, 92, 94, 95, 96, 97, 103, 104, 105, 106, 108, 125],
        [12, 13, 15, 16, 19, 20, 24, 38, 39, 52, 53, 68, 69, 70, 75, 98, 100, 101, 102, 106, 115, 116, 117, 125],
        [10, 15, 16, 17, 20, 21, 22, 24, 28, 45, 46, 47, 50, 60, 61, 62, 66, 85, 88, 89, 93, 105, 127, 128, 129],
        [15, 16, 21, 22, 24, 45, 47, 53, 56, 57, 58, 60, 61, 62, 66, 70, 98, 103, 105, 106, 108, 116, 117, 126],
        [4, 5, 10, 11, 12, 13, 15, 16, 30, 31, 32, 36, 38, 71, 73, 74, 94, 95, 96, 97, 109, 110, 111, 118, 120],
        [4, 9, 26, 27, 29, 35, 36, 37, 38, 53, 70, 71, 80, 82, 83, 97, 98, 106, 110, 111, 115, 116, 117],
        [15, 16, 21, 22, 24, 45, 47, 54, 55, 56, 60, 61, 62, 66, 93, 98, 103, 104, 105, 106, 108, 126],
        [14, 19, 20, 24, 38, 40, 41, 42, 44, 48, 49, 77, 112, 113, 114, 121, 122, 125, 127, 128, 129],
        [4, 5, 27, 28, 29, 33, 34, 35, 36, 60, 71, 73, 74, 92, 94, 95, 96, 97, 105, 121, 122, 125],
        [4, 5, 29, 33, 34, 35, 36, 37, 38, 71, 74, 80, 81, 82, 83, 95, 96, 97, 109, 110, 111, 118],
        [2, 3, 14, 18, 19, 20, 21, 22, 24, 28, 60, 64, 77, 79, 96, 97, 105, 114, 121, 122, 125],
        [2, 3, 14, 18, 19, 20, 21, 22, 24, 28, 60, 64, 71, 77, 78, 79, 96, 97, 105, 122, 125],
        [1, 6, 7, 8, 29, 33, 34, 35, 36, 37, 38, 71, 77, 78, 79, 96, 97, 109, 110, 111, 118],
        [1, 28, 29, 33, 34, 35, 36, 60, 71, 76, 77, 80, 96, 97, 105, 114, 125]
    ]

    print("=" * 70)
    print("baseline PSO")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("Program completed")


if __name__ == "__main__":
    main()