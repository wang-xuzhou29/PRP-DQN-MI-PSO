
import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
import numpy as np
import random
import time
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# ===  ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'MIN_X': 1, 'MAX_X': 100,
    'MIN_Y': 1, 'MAX_Y': 100,
    'MIN_Z': 1, 'MAX_Z': 100,
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 5,
    'NUM_ROUNDS': 5,
    'REPLAY_BATCH_SIZE': 64,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'LEARNING_RATE': 3e-4,
    'NUM_RUNS': 20,  # 20 run
    'TOP_K_SAMPLES': 20,
    'REPLAY_BUFFER_CAPACITY': 20000,  # Path 
    'TARGET_PATHS': [
        [15, 16, 19, 20, 21, 22, 24, 42, 48, 49, 51, 52, 53, 59, 60, 64, 70, 75, 84, 85, 86, 88, 89, 93, 98, 100, 101,
         102,
         103, 105, 106, 108, 116, 117],
        [15, 16, 23, 26, 27, 42, 48, 49, 51, 52, 53, 59, 60, 64, 70, 75, 84, 85, 86, 88, 89, 93, 98, 100, 101, 102, 103,
         105, 106, 108, 116, 117],
        [15, 16, 19, 20, 21, 22, 24, 42, 48, 49, 51, 52, 53, 59, 60, 64, 70, 75, 90, 91, 92, 98, 100, 101, 102, 103,
         105,
         106, 108, 116, 117],
        [13, 15, 16, 19, 20, 21, 22, 24, 42, 52, 53, 60, 64, 69, 70, 75, 91, 92, 98, 100, 101, 102, 103, 105, 106, 108,
         116,
         117, 123, 124],
        [12, 13, 15, 16, 19, 20, 21, 22, 24, 52, 53, 60, 64, 68, 69, 70, 75, 91, 92, 98, 100, 101, 102, 103, 105, 106,
         108,
         116, 117, 124],
        [11, 12, 13, 15, 16, 18, 19, 20, 21, 22, 24, 52, 53, 60, 64, 67, 70, 71, 92, 97, 98, 99, 103, 105, 106, 108,
         116,
         117, 123, 124],
        [15, 16, 20, 22, 24, 45, 46, 47, 50, 51, 52, 53, 61, 65, 70, 84, 85, 86, 88, 89, 93, 98, 102, 106, 107, 115,
         116,
         117, 128, 129],
        [11, 12, 13, 15, 16, 18, 19, 20, 21, 22, 24, 52, 53, 60, 64, 71, 72, 73, 74, 92, 97, 98, 99, 103, 104, 105, 106,
         108, 123, 124],
        [11, 12, 13, 15, 16, 18, 19, 20, 21, 22, 24, 60, 64, 71, 72, 73, 74, 92, 97, 98, 99, 103, 104, 105, 106, 108,
         121,
         122, 125],
        [15, 16, 20, 22, 24, 45, 47, 50, 51, 52, 53, 61, 65, 70, 84, 85, 86, 88, 89, 93, 98, 102, 106, 115, 116, 117,
         127,
         128, 129],
        [15, 16, 20, 38, 40, 41, 42, 44, 45, 47, 50, 51, 52, 53, 70, 75, 98, 101, 102, 106, 110, 111, 115, 116, 117,
         125,
         128, 129],
        [4, 11, 12, 13, 15, 16, 26, 27, 30, 31, 32, 36, 53, 60, 71, 72, 73, 74, 92, 97, 98, 103, 104, 105, 106, 108,
         121,
         122, 125],
        [15, 16, 20, 21, 22, 24, 45, 47, 51, 52, 53, 60, 61, 62, 66, 70, 93, 98, 102, 103, 105, 106, 108, 116, 117, 127,
         128, 129],
        [15, 16, 20, 38, 40, 41, 42, 44, 45, 46, 47, 50, 51, 52, 53, 70, 75, 98, 101, 102, 106, 110, 111, 117, 118, 125,
         128, 129],
        [13, 15, 16, 25, 42, 51, 52, 53, 59, 60, 64, 69, 70, 75, 85, 86, 88, 89, 93, 98, 100, 101, 102, 103, 105, 106,
         108,
         117],
        [15, 16, 20, 21, 22, 24, 48, 49, 51, 52, 53, 59, 60, 63, 64, 85, 88, 89, 93, 98, 101, 102, 103, 104, 105, 106,
         108,
         129],
        [15, 16, 20, 38, 40, 41, 42, 44, 45, 46, 47, 50, 51, 52, 53, 98, 101, 102, 106, 110, 111, 118, 119, 120, 125,
         128,
         129],
        [13, 15, 16, 42, 51, 52, 53, 59, 60, 64, 69, 70, 75, 87, 90, 91, 92, 98, 100, 101, 102, 103, 105, 106, 108, 116,
         117],
        [15, 16, 19, 20, 24, 42, 43, 48, 49, 51, 52, 53, 70, 75, 92, 98, 100, 101, 102, 106, 115, 116, 117, 123, 124],
        [4, 5, 10, 11, 12, 13, 15, 16, 27, 28, 60, 71, 73, 74, 92, 94, 95, 96, 97, 103, 104, 105, 106, 108, 123, 124],
        [15, 16, 23, 26, 27, 45, 47, 53, 56, 57, 58, 60, 61, 62, 70, 98, 103, 105, 106, 108, 116, 117, 127, 128, 129],
        [4, 5, 27, 28, 29, 34, 35, 36, 60, 71, 73, 74, 80, 82, 83, 92, 94, 95, 96, 97, 103, 104, 105, 106, 108, 125],
        [12, 13, 15, 16, 19, 20, 24, 38, 39, 52, 53, 68, 69, 70, 75, 98, 100, 101, 102, 106, 115, 116, 117, 125],
        [10, 15, 16, 17, 20, 21, 22, 24, 28, 45, 46, 47, 50, 60, 61, 62, 66, 85, 88, 89, 93, 105, 127, 128, 129],
        [15, 16, 21, 22, 24, 45, 47, 53, 56, 57, 58, 60, 61, 62, 66, 70, 98, 103, 105, 106, 108, 116, 117, 126],
        [4, 5, 10, 11, 12, 13, 15, 16, 30, 31, 32, 36, 38, 71, 73, 74, 94, 95, 96, 97, 109, 110, 111, 118, 120],
        [4, 9, 26, 27, 29, 35, 36, 37, 38, 53, 70, 71, 80, 82, 83, 97, 98, 106, 110, 111, 115, 116, 117],
        [15, 16, 21, 22, 24, 45, 47, 54, 55, 56, 60, 61, 62, 66, 93, 98, 103, 104, 105, 106, 108, 126],
        [14, 19, 20, 24, 38, 40, 41, 42, 44, 48, 49, 77, 112, 113, 114, 121, 122, 125, 127, 128, 129],
        [4, 5, 27, 28, 29, 33, 34, 35, 36, 60, 71, 73, 74, 92, 94, 95, 96, 97, 105, 121, 122, 125],
        [4, 5, 29, 33, 34, 35, 36, 37, 38, 71, 74, 80, 81, 82, 83, 95, 96, 97, 109, 110, 111, 118],
        [2, 3, 14, 18, 19, 20, 21, 22, 24, 28, 60, 64, 77, 79, 96, 97, 105, 114, 121, 122, 125],
        [2, 3, 14, 18, 19, 20, 21, 22, 24, 28, 60, 64, 71, 77, 78, 79, 96, 97, 105, 122, 125],
        [1, 6, 7, 8, 29, 33, 34, 35, 36, 37, 38, 71, 77, 78, 79, 96, 97, 109, 110, 111, 118],
        [1, 28, 29, 33, 34, 35, 36, 60, 71, 76, 77, 80, 96, 97, 105, 114, 125]
    ],
}


# ===  ===
def get_bounds():
    mins = np.array([EXPERIMENT_CONFIG['MIN_X'], EXPERIMENT_CONFIG['MIN_Y'], EXPERIMENT_CONFIG['MIN_Z']])
    maxs = np.array([EXPERIMENT_CONFIG['MAX_X'], EXPERIMENT_CONFIG['MAX_Y'], EXPERIMENT_CONFIG['MAX_Z']])
    return mins, maxs

def clip_state(state):
    mins, maxs = get_bounds()
    return np.clip(state, mins, maxs)

def denormalize_state(normalized_state):
    """"""
    mins, maxs = get_bounds()
    return normalized_state * (maxs - mins) / 2 + (mins + maxs) / 2


def coverage_similarity(triggered, target_path):
    """
    Similarity: / target paths
    """
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0

    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)


def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']

    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']

    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']

    return reward


# ===   ===
def execute_Tr(light, moisture, co2):
    actions = []
    triggered = set()

    # Fixed all if statements - using triggered.add() instead of b[0]=1
    if (light < 3500 and moisture > 75) != (light < 2500 and moisture > 75):
        triggered.add(1)
    if (light < 3500 and moisture > 75) != (light < 4500 and moisture > 75):
        triggered.add(2)
    if (light < 3500 and moisture > 75) != (light < 5500 and moisture > 75):
        triggered.add(3)
    if (light < 3500 and moisture > 75) != (light < 3500 and moisture > 25):
        triggered.add(4)
    if (light < 3500 and moisture > 75) != (light < 3500 and moisture > 55):
        triggered.add(5)

    if (light < 2500 and moisture > 80 and co2 > 1600) != (light < 5500 and moisture > 80 and co2 > 1600):
        triggered.add(6)
    if (light < 2500 and moisture > 80 and co2 > 1600) != (light < 6500 and moisture > 80 and co2 > 1600):
        triggered.add(7)
    if (light < 2500 and moisture > 80 and co2 > 1600) != (light < 8500 and moisture > 80 and co2 > 1600):
        triggered.add(8)
    if (light < 2500 and moisture > 80 and co2 > 1600) != (light < 2500 and moisture > 40 and co2 > 1600):
        triggered.add(9)
    # 原分支10已删除

    if (light > 3000 and moisture < 75) != (light > 3000 and moisture < 55):
        triggered.add(10)
    if (light > 3000 and moisture < 75) != (light > 4000 and moisture < 75):
        triggered.add(11)
    if (light > 3000 and moisture < 75) != (light > 5000 and moisture < 75):
        triggered.add(12)
    if (light > 3000 and moisture < 75) != (light > 6000 and moisture < 75):
        triggered.add(13)
    if (light > 3000 and moisture < 75) != (light > 3000 and moisture < 85):
        triggered.add(14)
    if (light > 3000 and moisture < 75) != (light > 3000 and moisture < 35):
        triggered.add(15)
    if (light > 3000 and moisture < 75) != (light > 3000 and moisture < 15):
        triggered.add(16)
    if (light > 3000 and moisture < 75) != (light > 3000 and moisture < 65):
        triggered.add(17)

    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 4500 or co2 < 800) and moisture > 42):
        triggered.add(18)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 6500 or co2 < 800) and moisture > 42):
        triggered.add(19)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 8500 or co2 < 800) and moisture > 42):
        triggered.add(20)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 3500 or co2 < 900) and moisture > 42):
        triggered.add(21)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 3500 or co2 < 1100) and moisture > 42):
        triggered.add(22)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 3500 or co2 < 700) and moisture > 42):
        triggered.add(23)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 3500 or co2 < 1500) and moisture > 42):
        triggered.add(24)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 3500 or co2 < 800) and moisture > 32):
        triggered.add(25)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 3500 or co2 < 800) and moisture > 52):
        triggered.add(26)
    if ((light < 3500 or co2 < 800) and moisture > 42) != ((light < 3500 or co2 < 800) and moisture > 62):
        triggered.add(27)

    if (moisture / (light / 100) > 8) != (moisture / (co2 / 100) > 8):
        triggered.add(28)

    if (light < 3000 and (light + moisture * 10) < 3800) != (light < 2000 and (light + moisture * 10) < 3800):
        triggered.add(29)
    if (light < 3000 and (light + moisture * 10) < 3800) != (light < 4000 and (light + moisture * 10) < 3800):
        triggered.add(30)
    if (light < 3000 and (light + moisture * 10) < 3800) != (light < 5000 and (light + moisture * 10) < 3800):
        triggered.add(31)
    if (light < 3000 and (light + moisture * 10) < 3800) != (light < 7000 and (light + moisture * 10) < 3800):
        triggered.add(32)
    if (light < 3000 and (light + moisture * 10) < 3800) != (light < 3000 and (light + moisture * 15) < 3800):
        triggered.add(33)
    if (light < 3000 and (light + moisture * 10) < 3800) != (light < 3000 and (light + moisture * 20) < 3800):
        triggered.add(34)
    if (light < 3000 and (light + moisture * 10) < 3800) != (light < 3000 and (light + moisture * 10) > 3800):
        triggered.add(35)
    if (light < 3000 and (light + moisture * 10) < 3800) != (light > 3000 and (light + moisture * 10) < 3800):
        triggered.add(36)
    # 原分支38已删除
    if (light < 3000 and (light + moisture * 10) < 3800) != (light < 3000 and (light + co2) < 3800):
        triggered.add(37)

    if (light > 5000 and light < 7000 and co2 > 1400) != (light < 5000 and light < 7000 and co2 > 1400):
        triggered.add(38)
    if (light > 5000 and light < 7000 and co2 > 1400) != (light > 4000 and light < 7000 and co2 > 1400):
        triggered.add(39)
    if (light > 5000 and light < 7000 and co2 > 1400) != (light > 5000 and light > 7000 and co2 > 1400):
        triggered.add(40)
    if (light > 5000 and light < 7000 and co2 > 1400) != (light > 5000 and light < 3000 and co2 > 1400):
        triggered.add(41)
    if (light > 5000 and light < 7000 and co2 > 1400) != (light > 5000 and light < 7000 and co2 < 1400):
        triggered.add(42)
    if (light > 5000 and light < 7000 and co2 > 1400) != (light > 5000 and light < 7000 and co2 > 1200):
        triggered.add(43)
    if (light > 5000 and light < 7000 and co2 > 1400) != (light > 5000 and light < 4000 and co2 > 1400):
        triggered.add(44)

    if (light > 6000 and (light + moisture + co2) > 8500) != (light < 6000 and (light + moisture + co2) > 8500):
        triggered.add(45)
    if (light > 6000 and (light + moisture + co2) > 8500) != (light > 6000 and (light - moisture + co2) > 8500):
        triggered.add(46)
    if (light > 6000 and (light + moisture + co2) > 8500) != (light > 6000 and (light + moisture - co2) > 8500):
        triggered.add(47)
    if (light > 6000 and (light + moisture + co2) > 8500) != (light > 6000 and (light + moisture + co2) > 4500):
        triggered.add(48)
    if (light > 6000 and (light + moisture + co2) > 8500) != (light > 6000 and (light + moisture + co2) > 6500):
        triggered.add(49)
    if (light > 6000 and (light + moisture + co2) > 8500) != (light > 8000 and (light + moisture + co2) > 8500):
        triggered.add(50)

    if (light > 8500 and moisture < 50) != (light > 5500 and moisture < 50):
        triggered.add(51)
    if (light > 8500 and moisture < 50) != (light > 3500 and moisture < 50):
        triggered.add(52)
    if (light > 8500 and moisture < 50) != (light < 8500 and moisture < 50):
        triggered.add(53)
    if (light > 8500 and moisture < 50) != (light > 8500 and moisture < 70):
        triggered.add(54)
    if (light > 8500 and moisture < 50) != (light > 8500 and moisture < 60):
        triggered.add(55)
    if (light > 8500 and moisture < 50) != (light > 8500 and moisture > 50):
        triggered.add(56)
    if (light > 8500 and moisture < 50) != (light > 8500 and moisture < 40):
        triggered.add(57)
    if (light > 8500 and moisture < 50) != (light > 8500 and moisture < 30):
        triggered.add(58)

    if (light > 7500 and co2 < 900) != (light > 5500 and co2 < 900):
        triggered.add(59)
    if (light > 7500 and co2 < 900) != (light < 7500 and co2 < 900):
        triggered.add(60)
    if (light > 7500 and co2 < 900) != (light > 7500 and co2 > 900):
        triggered.add(61)
    if (light > 7500 and co2 < 900) != (light > 7500 and co2 < 400):
        triggered.add(62)
    if (light > 7500 and co2 < 900) != (light > 7000 and co2 < 900):
        triggered.add(63)
    if (light > 7500 and co2 < 900) != (light > 3500 and co2 < 900):
        triggered.add(64)
    if (light > 7500 and co2 < 900) != (light > 7500 and co2 < 1400):
        triggered.add(65)
    if (light > 7500 and co2 < 900) != (light > 7500 and co2 < 800):
        triggered.add(66)

    if (light < 4000 and moisture < 45) != (light < 3000 and moisture < 45):
        triggered.add(67)
    if (light < 4000 and moisture < 45) != (light < 5000 and moisture < 45):
        triggered.add(68)
    if (light < 4000 and moisture < 45) != (light < 6000 and moisture < 45):
        triggered.add(69)
    if (light < 4000 and moisture < 45) != (light > 4000 and moisture < 45):
        triggered.add(70)
    if (light < 4000 and moisture < 45) != (light < 4000 and moisture > 45):
        triggered.add(71)
    if (light < 4000 and moisture < 45) != (light < 4000 and moisture < 55):
        triggered.add(72)
    if (light < 4000 and moisture < 45) != (light < 4000 and moisture < 65):
        triggered.add(73)
    if (light < 4000 and moisture < 45) != (light < 4000 and moisture < 75):
        triggered.add(74)
    if (light < 4000 and moisture < 45) != (light < 7000 and moisture < 45):
        triggered.add(75)

    if (light < 2800 and moisture > 80) != (light < 1800 and moisture > 80):
        triggered.add(76)
    if (light < 2800 and moisture > 80) != (light > 2800 and moisture > 80):
        triggered.add(77)
    if (light < 2800 and moisture > 80) != (light < 3800 and moisture > 80):
        triggered.add(78)
    if (light < 2800 and moisture > 80) != (light < 4800 and moisture > 80):
        triggered.add(79)
    if (light < 2800 and moisture > 80) != (light < 2800 and moisture < 80):
        triggered.add(80)
    if (light < 2800 and moisture > 80) != (light < 2800 and moisture > 60):
        triggered.add(81)
    if (light < 2800 and moisture > 80) != (light < 2800 and moisture > 40):
        triggered.add(82)
    if (light < 2800 and moisture > 80) != (light < 2800 and moisture > 30):
        triggered.add(83)

    if (light / (co2 + 1) > 7) != (light / (co2 + 100) > 7):
        triggered.add(84)
    if (light / (co2 + 1) > 7) != (light / (co2 + 200) > 7):
        triggered.add(85)
    if (light / (co2 + 1) > 7) != (light / (co2 + 1) > 8):
        triggered.add(86)
    if (light / (co2 + 1) > 7) != (light / (co2 - 1) > 7):
        triggered.add(87)
    if (light / (co2 + 1) > 7) != (light / (co2 + 1) > 9):
        triggered.add(88)
    if (light / (co2 + 1) > 7) != (light / (co2 + 300) > 7):
        triggered.add(89)
    if (light / (co2 + 1) > 7) != (light / (co2 - 100) > 7):
        triggered.add(90)
    if (light / (co2 + 1) > 7) != (light / (co2 - 200) > 7):
        triggered.add(91)
    if (light / (co2 + 1) > 7) != (light / (co2 - 300) > 7):
        triggered.add(92)
    if (light / (co2 + 1) > 7) != (light / (co2 + 1) > 10):
        triggered.add(93)

    if (moisture < 55 and light < 4500) != (moisture < 65 and light < 4500):
        triggered.add(94)
    if (moisture < 55 and light < 4500) != (moisture < 75 and light < 4500):
        triggered.add(95)
    if (moisture < 55 and light < 4500) != (moisture < 85 and light < 4500):
        triggered.add(96)
    if (moisture < 55 and light < 4500) != (moisture > 55 and light < 4500):
        triggered.add(97)
    if (moisture < 55 and light < 4500) != (moisture < 55 and light > 4500):
        triggered.add(98)
    if (moisture < 55 and light < 4500) != (moisture < 55 and light < 3500):
        triggered.add(99)
    if (moisture < 55 and light < 4500) != (moisture < 55 and light < 6500):
        triggered.add(100)
    if (moisture < 55 and light < 4500) != (moisture < 55 and light < 7500):
        triggered.add(101)
    if (moisture < 55 and light < 4500) != (moisture < 55 and light < 8500):
        triggered.add(102)

    if (moisture < 58 and co2 < 950) != (moisture < 38 and co2 < 950):
        triggered.add(103)
    if (moisture < 58 and co2 < 950) != (moisture < 48 and co2 < 950):
        triggered.add(104)
    if (moisture < 58 and co2 < 950) != (moisture > 58 and co2 < 950):
        triggered.add(105)
    if (moisture < 58 and co2 < 950) != (moisture < 58 and co2 > 950):
        triggered.add(106)
    if (moisture < 58 and co2 < 950) != (moisture < 58 and co2 < 1050):
        triggered.add(107)
    if (moisture < 58 and co2 < 950) != (moisture < 58 and co2 < 450):
        triggered.add(108)

    if (moisture > 82 and co2 > 1600) != (moisture > 62 and co2 > 1600):
        triggered.add(109)
    if (moisture > 82 and co2 > 1600) != (moisture > 32 and co2 > 1600):
        triggered.add(110)
    if (moisture > 82 and co2 > 1600) != (moisture > 42 and co2 > 1600):
        triggered.add(111)
    if (moisture > 82 and co2 > 1600) != (moisture > 82 and co2 > 1100):
        triggered.add(112)
    if (moisture > 82 and co2 > 1600) != (moisture > 82 and co2 > 1000):
        triggered.add(113)
    if (moisture > 82 and co2 > 1600) != (moisture > 82 and co2 > 700):
        triggered.add(114)

    if (co2 > 1750 and moisture < 45) != (co2 > 950 and moisture < 45):
        triggered.add(115)
    if (co2 > 1750 and moisture < 45) != (co2 > 750 and moisture < 45):
        triggered.add(116)
    if (co2 > 1750 and moisture < 45) != (co2 < 1750 and moisture < 45):
        triggered.add(117)
    if (co2 > 1750 and moisture < 45) != (co2 > 1750 and moisture > 45):
        triggered.add(118)
    if (co2 > 1750 and moisture < 45) != (co2 > 1750 and moisture < 55):
        triggered.add(119)
    if (co2 > 1750 and moisture < 45) != (co2 > 1750 and moisture < 65):
        triggered.add(120)

    if (co2 / (light + 1) > 0.22) != (co2 / (light + 100) > 0.22):
        triggered.add(121)
    if (co2 / (light + 1) > 0.22) != (co2 / (light + 200) > 0.22):
        triggered.add(122)
    if (co2 / (light + 1) > 0.22) != (co2 / (light - 1000) > 0.22):
        triggered.add(123)
    if (co2 / (light + 1) > 0.22) != (co2 / (light - 2000) > 0.22):
        triggered.add(124)
    if (co2 / (light + 1) > 0.22) != (co2 / (light + 1) > 0.32):
        triggered.add(125)

    if (light + moisture * 25 + co2 > 10500) != (light + moisture * 15 + co2 > 10500):
        triggered.add(126)
    if (light + moisture * 25 + co2 > 10500) != (light + moisture * 35 + co2 > 10500):
        triggered.add(127)
    if (light + moisture * 25 + co2 > 10500) != (light + moisture * 45 + co2 > 10500):
        triggered.add(128)
    if (light + moisture * 25 + co2 > 10500) != (light + moisture * 25 + co2 + 1000 > 10500):
        triggered.add(129)

    return triggered



# 
execute_Tr = execute_Tr

# === DQN ===
class DQNNetwork(nn.Module):
    def __init__(self, action_size=30):
        super(DQNNetwork, self).__init__()
        hidden_dim = EXPERIMENT_CONFIG['HIDDEN_DIM']

        self.fc1 = nn.Linear(3, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)
        self.output = nn.Linear(hidden_dim, action_size)

    def forward(self, x):
        x = torch.relu(self.fc1(x))
        x = torch.relu(self.fc2(x))
        x = torch.relu(self.fc3(x))
        return self.output(x)


# === Path  ===
class PathReplayBuffer:
    """Path """

    def __init__(self, path_idx, capacity=20000):
        self.path_idx = path_idx
        self.capacity = capacity
        self.buffer = deque(maxlen=capacity)
        self.similarities = deque(maxlen=capacity)  # Similarity

    def push(self, state, action, reward, next_state, done, similarity):
        self.buffer.append((state, action, reward, next_state, done))
        self.similarities.append(similarity)

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return None

        batch = random.sample(self.buffer, batch_size)
        state, action, reward, next_state, done = map(np.stack, zip(*batch))

        return (
            torch.FloatTensor(state).to(device),
            torch.LongTensor(action).to(device),
            torch.FloatTensor(reward).to(device),
            torch.FloatTensor(next_state).to(device),
            torch.BoolTensor(done).to(device)
        )

    def get_top_k(self, k=20):
        """Path Top-K"""
        if len(self.buffer) == 0:
            return []

        # buffersimilarities
        samples_with_sim = list(zip(self.buffer, self.similarities))
        samples_with_sim.sort(key=lambda x: x[1], reverse=True)

        # Top-K
        top_k = samples_with_sim[:k]

        results = []
        for (state, _, _, _, _), similarity in top_k:
            original_state = denormalize_state(state)
            original_state_int = np.round(original_state).astype(int)

            results.append({
                'state': original_state_int,
                'similarity': similarity,
                'triggered': execute_Tr(*original_state_int)
            })

        return results

    def __len__(self):
        return len(self.buffer)


# === DQN(:)===
class ImprovedDQNAgent:
    def __init__(self, num_paths, action_size=30):
        self.action_size = action_size
        self.num_paths = num_paths
        self.epsilon = 0.9
        self.epsilon_min = 0.1
        self.epsilon_decay = 0.995

        self.q_network = DQNNetwork(action_size).to(device)
        self.target_network = DQNNetwork(action_size).to(device)

        lr = EXPERIMENT_CONFIG['LEARNING_RATE']
        self.optimizer = optim.Adam(self.q_network.parameters(), lr=lr)

        # Path 
        capacity = EXPERIMENT_CONFIG['REPLAY_BUFFER_CAPACITY']
        self.replay_buffers = {}
        for path_idx in range(num_paths):
            self.replay_buffers[path_idx] = PathReplayBuffer(path_idx, capacity)

        self.replay_train_count = 0
        self.update_target_network()

    def discrete_to_action_delta(self, action_idx):
        # 
        delta_values = [5, 3, 2, 1, 0.5, -0.5, -1, -2, -3, -5]

        if action_idx >= 30:
            action_idx = action_idx % 30

        dim = action_idx // 10
        delta_idx = action_idx % 10
        delta = delta_values[delta_idx]

        action_delta = np.zeros(3)
        action_delta[dim] = delta

        return action_delta

    def get_action(self, state):
        mins, maxs = get_bounds()
        normalized_state = (state - (mins + maxs) / 2) / ((maxs - mins) / 2)

        if random.random() < self.epsilon:
            action_idx = random.randint(0, self.action_size - 1)
        else:
            state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)
            with torch.no_grad():
                q_values = self.q_network(state_tensor)
                action_idx = q_values.argmax().item()

        action_delta = self.discrete_to_action_delta(action_idx)
        return action_delta, action_idx

    def store_experience(self, path_idx, state, action_idx, reward, next_state, done, similarity):
        """Path """
        mins, maxs = get_bounds()
        normalized_state = (state - (mins + maxs) / 2) / ((maxs - mins) / 2)
        normalized_next_state = (next_state - (mins + maxs) / 2) / ((maxs - mins) / 2)

        self.replay_buffers[path_idx].push(
            normalized_state, action_idx, reward,
            normalized_next_state, done, similarity
        )

    def replay_train(self, path_idx):
        """Path """
        batch_size = EXPERIMENT_CONFIG['REPLAY_BATCH_SIZE']
        batch = self.replay_buffers[path_idx].sample(batch_size)

        if batch is None:
            return

        states, actions, rewards, next_states, dones = batch

        current_q_values = self.q_network(states).gather(1, actions.unsqueeze(1))

        with torch.no_grad():
            next_q_values = self.target_network(next_states).max(1)[0]
            target_q_values = rewards + (0.99 * next_q_values * ~dones)

        loss = F.mse_loss(current_q_values.squeeze(), target_q_values)

        self.optimizer.zero_grad()
        loss.backward()
        torch.nn.utils.clip_grad_norm_(self.q_network.parameters(), 1.0)
        self.optimizer.step()

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

        self.replay_train_count += 1

        if self.replay_train_count % 2 == 0:
            self.update_target_network()
            print(f"    ->  (Run {self.replay_train_count})")

    def update_target_network(self):
        self.target_network.load_state_dict(self.q_network.state_dict())

    def get_all_top_k(self, k=20):
        """Path Top-K"""
        results = {}
        for path_idx in range(self.num_paths):
            results[path_idx] = self.replay_buffers[path_idx].get_top_k(k)
        return results

    def get_buffer_stats(self):
        """"""
        stats = {}
        for path_idx in range(self.num_paths):
            stats[path_idx] = len(self.replay_buffers[path_idx])
        return stats


# === Metric ===
def calculate_run_performance(run_idx, dqn_results, training_time, total_steps, update_count, agent):
    """ runMetric"""
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    # 1. (Total Reward)
    total_reward = 0
    # 2. (Average Reward)
    average_reward = 0
    # 5. (Convergence)
    convergence = 0
    # 12. (Environment Adaptability)
    environment_adaptability = 0
    # 13. (Generalization Ability)
    generalization_ability = 0
    # 15. (Computational Efficiency)
    computational_efficiency = 0
    # 16. (Policy Update Frequency)
    policy_update_frequency = 0

    # Similarity
    all_similarities = []

    # Metric
    total_samples = 0
    all_rewards = []

    for path_idx in range(num_paths):
        samples = dqn_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']

            total_reward += reward
            all_rewards.append(reward)
            all_similarities.append(similarity)
            total_samples += 1

    # 1. 
    total_reward = total_reward

    # 2. 
    if total_samples > 0:
        average_reward = total_reward / total_samples

    # 5. (Average Similarity)
    if all_similarities:
        convergence = np.mean(all_similarities)

    # 12. (Similarity)
    if len(all_similarities) > 1:
        environment_adaptability = 1 / (np.std(all_similarities) + 1e-8)

    # 13. (Average Similarity)
    generalization_ability = convergence

    # 15. (/ seconds)
    if training_time > 0:
        computational_efficiency = total_steps / training_time

    # 16. 
    if training_time > 0:
        policy_update_frequency = update_count / training_time

    # Similarity
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0

    return {
        'Run': run_idx + 1,

        # Metric
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),

        # Similarity
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }


# === Excel ===
def export_to_excel(all_dqn_results, all_performance_data, target_paths, output_path="DQN_20 run.xlsx"):
    """20 runDQNExcel"""
    print("\nExcel...")

    # 
    all_dqn_summary_data = []
    all_dqn_detailed_data = []

    #  run
    for run_idx, (dqn_results, performance_data) in enumerate(zip(all_dqn_results, all_performance_data)):
        # ===== Sheet1: DQNPath  =====
        dqn_summary_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = dqn_results[path_idx]

            if len(samples) == 0:
                dqn_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue

            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = 'Yes' if perfect_count > 0 else 'No'

            dqn_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })
        all_dqn_summary_data.extend(dqn_summary_data)

        # ===== Sheet2: DQNDetailed Sample Data =====
        dqn_detailed_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = dqn_results[path_idx]

            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']

                dqn_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Sample ID': sample_idx + 1,
                    'X': int(state[0]),
                    'Y': int(state[1]),
                    'Z': int(state[2]),
                    'Similarity': round(similarity, 4),
                    'Perfect Coverage': 'Yes' if abs(similarity - 1.0) < 0.001 else 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path))),
                    'Triggered Rules': ', '.join(map(str, sorted(triggered))),
                    'Hit Rule Count': len(target_path.intersection(triggered)),
                    'Target Rule Count': len(target_path)
                })
        all_dqn_detailed_data.extend(dqn_detailed_data)

    # Excel
    dqn_summary_df = pd.DataFrame(all_dqn_summary_data)
    dqn_detailed_df = pd.DataFrame(all_dqn_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet1: DQNPath 
        dqn_summary_df.to_excel(writer, sheet_name='DQNPath ', index=False)

        # Sheet2: DQNDetailed Sample Data
        dqn_detailed_df.to_excel(writer, sheet_name='DQNDetailed Sample Data', index=False)

        # Sheet3: Metric
        selected_columns = [
            'Run',
            'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df_selected = performance_df[selected_columns]
        performance_df_selected.to_excel(writer, sheet_name='Metric', index=False)

        # 
        workbook = writer.book

        # 
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        perfect_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # 

        # === Sheet1 ===
        ws1 = writer.sheets['DQNPath ']
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 
        for row_idx in range(2, ws1.max_row + 1):
            if ws1.cell(row_idx, 9).value == '':  # Run 9""
                for col_idx in range(1, ws1.max_column + 1):
                    ws1.cell(row_idx, col_idx).fill = perfect_fill

        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 15
        ws1.column_dimensions['F'].width = 15
        ws1.column_dimensions['G'].width = 15
        ws1.column_dimensions['H'].width = 15
        ws1.column_dimensions['I'].width = 15
        ws1.column_dimensions['J'].width = 50

        # === Sheet2 ===
        ws2 = writer.sheets['DQNDetailed Sample Data']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 15
        ws2.column_dimensions['I'].width = 40
        ws2.column_dimensions['J'].width = 40
        ws2.column_dimensions['K'].width = 15
        ws2.column_dimensions['L'].width = 15

        # === Sheet3 ===
        ws3 = writer.sheets['Metric']
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        for col in columns:
            ws3.column_dimensions[col].width = 18

    print(f"Excel: {output_path}")
    print(f"  - Sheet1: DQNPath  ({len(all_dqn_summary_data)})")
    print(f"  - Sheet2: DQNDetailed Sample Data ({len(all_dqn_detailed_data)})")
    print(f"  - Sheet3: Metric ({len(all_performance_data)})")


# === DQN training(:)===
def train_dqn_workflow():
    print("=" * 80)
    print("DQN training ()")
    print("Similarity:  / target paths")
    print(
        f": Path {EXPERIMENT_CONFIG['NUM_ROUNDS']},{EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(f": Path ,={EXPERIMENT_CONFIG['REPLAY_BUFFER_CAPACITY']}")
    print("=" * 80)

    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    # (Number of Paths)
    agent = ImprovedDQNAgent(num_paths=num_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n: Path {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']}")
    path_samples = {}
    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            state = np.array([
                random.randint(EXPERIMENT_CONFIG['MIN_X'], EXPERIMENT_CONFIG['MAX_X']),
                random.randint(EXPERIMENT_CONFIG['MIN_Y'], EXPERIMENT_CONFIG['MAX_Y']),
                random.randint(EXPERIMENT_CONFIG['MIN_Z'], EXPERIMENT_CONFIG['MAX_Z'])
            ], dtype=np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  Path  {path_idx + 1}/{num_paths}:  {len(samples)} ")

    # 
    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size
    num_rounds = EXPERIMENT_CONFIG['NUM_ROUNDS']

    print(f"\n:")
    print(f"  - : {batch_size}")
    print(f"  - Path : {num_batches}")
    print(f"  - Path : {num_rounds}")
    print(f"  - : {EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(
        f"  - : {num_paths} Path  x {num_rounds}  x {num_batches}  = {num_paths * num_rounds * num_batches} ")
    print("-" * 80)

    # :completedPath ,Path 
    for path_idx in range(num_paths):
        target_path = target_paths[path_idx]
        print(f"\n{'=' * 80}")
        print(f"Start training path  {path_idx + 1}/{num_paths}")
        print(f": {sorted(target_path)}")
        print(f": replay_buffers[{path_idx}]")
        print(f"{'=' * 80}")

        # Path NUM_ROUNDS
        for round_idx in range(num_rounds):
            print(f"\n{'' * 80}")
            print(f"Path  {path_idx + 1} - Run  {round_idx + 1}/{num_rounds} ")
            print(f"{'' * 80}")

            # Per roundnum_batches
            for batch_idx in range(num_batches):
                print(f"\n   {batch_idx + 1}/{num_batches} (Path {path_idx + 1}, Run {round_idx + 1})")

                # 
                batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]

                batch_rewards = []
                batch_similarities = []

                # 
                for sample_idx, initial_state in enumerate(batch_samples):
                    state = initial_state.copy()
                    episode_reward = 0
                    final_similarity = 0

                    # STEPS_PER_SAMPLE
                    for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                        action_delta, action_idx = agent.get_action(state)

                        next_state = state + action_delta
                        next_state = clip_state(next_state)

                        triggered = execute_Tr(*next_state)  #
                        reward = unified_reward_function(triggered, target_path)
                        similarity = coverage_similarity(triggered, target_path)

                        done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)

                        # Path 
                        agent.store_experience(
                            path_idx, state, action_idx, reward, next_state, done, similarity
                        )

                        state = next_state
                        episode_reward += reward
                        final_similarity = similarity
                        total_steps += 1

                    batch_rewards.append(episode_reward)
                    batch_similarities.append(final_similarity)

                # 
                avg_reward = np.mean(batch_rewards)
                avg_similarity = np.mean(batch_similarities)
                max_similarity = np.max(batch_similarities)
                print(f"    ={avg_reward:.2f}, Average Similarity={avg_similarity:.4f}, "
                      f"Similarity={max_similarity:.4f}, epsilon={agent.epsilon:.3f}")

                # Path 
                print(f"    (Path {path_idx})...")
                agent.replay_train(path_idx)

                # Path 
                buffer_size = len(agent.replay_buffers[path_idx])
                print(f"    Path {path_idx}: {buffer_size}, : {agent.replay_train_count}")

    training_time = time.time() - start_time

    print("\n" + "=" * 80)
    print(f"DQN trainingcompleted! Total elapsed time: {training_time:.2f} seconds, : {total_steps}")
    print(f": {agent.replay_train_count}")
    print(f": {agent.replay_train_count // 2}")

    # Path 
    print("\nPath :")
    buffer_stats = agent.get_buffer_stats()
    for path_idx, size in buffer_stats.items():
        print(f"  Path {path_idx + 1}: {size} ")

    print("=" * 80)

    # Top-K
    print(f"\nPath SimilarityMaximum{EXPERIMENT_CONFIG['TOP_K_SAMPLES']}...")
    dqn_top_k_results = agent.get_all_top_k(EXPERIMENT_CONFIG['TOP_K_SAMPLES'])

    return agent, dqn_top_k_results, training_time, total_steps, agent.replay_train_count


# ===  ===
def main():
    print("\n" + "=" * 80)
    print("DQN - 20 run")
    print("Metric")
    print("=" * 80)

    all_dqn_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    # 20
    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'='*80}")
        print(f"Start run  {run_idx + 1}/{EXPERIMENT_CONFIG['NUM_RUNS']}  run")
        print(f"{'='*80}")

        # DQN training
        dqn_agent, dqn_results, training_time, total_steps, update_count = train_dqn_workflow()

        # Metric
        performance_data = calculate_run_performance(
            run_idx, dqn_results, training_time, total_steps, update_count, dqn_agent
        )

        # 
        all_dqn_results.append(dqn_results)
        all_performance_data.append(performance_data)

        print(f"\nRun {run_idx + 1} completed!")
        print(f"  Total Reward: {performance_data['Total Reward']}")
        print(f"  Average Reward: {performance_data['Average Reward']}")
        print(f"  Convergence: {performance_data['Convergence']}")

    # Excel(20 run)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"DQN_20 run_{timestamp}.xlsx"
    export_to_excel(all_dqn_results, all_performance_data, target_paths, output_path)

    # 
    print("\n" + "=" * 80)
    print("20 run")
    print("=" * 80)

    # Metric
    # Metric Extraction
    total_rewards = [p['Total Reward'] for p in all_performance_data]
    average_rewards = [p['Average Reward'] for p in all_performance_data]
    convergences = [p['Convergence'] for p in all_performance_data]
    environment_adaptabilities = [p['Environment Adaptability'] for p in all_performance_data]
    generalization_abilities = [p['Generalization Ability'] for p in all_performance_data]
    computational_efficiencies = [p['Computational Efficiency'] for p in all_performance_data]
    policy_update_frequencies = [p['Policy Update Frequency'] for p in all_performance_data]
    avg_similarities = [p['Average Similarity'] for p in all_performance_data]

    print(f":")
    print(f"  : {np.mean(total_rewards):.2f}")
    print(f"  Standard deviation: {np.std(total_rewards):.2f}")

    print(f"\n:")
    print(f"  : {np.mean(average_rewards):.4f}")
    print(f"  Standard deviation: {np.std(average_rewards):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(convergences):.4f}")
    print(f"  Standard deviation: {np.std(convergences):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(environment_adaptabilities):.4f}")
    print(f"  Standard deviation: {np.std(environment_adaptabilities):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(generalization_abilities):.4f}")
    print(f"  Standard deviation: {np.std(generalization_abilities):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(computational_efficiencies):.2f}")
    print(f"  Standard deviation: {np.std(computational_efficiencies):.2f}")

    print(f"\n:")
    print(f"  : {np.mean(policy_update_frequencies):.4f}")
    print(f"  Standard deviation: {np.std(policy_update_frequencies):.4f}")

    print(f"\nAverage similarity statistics:")
    print(f"  : {np.mean(avg_similarities):.4f}")
    print(f"  Standard deviation: {np.std(avg_similarities):.4f}")

    print("\n" + "=" * 80)
    print(f" {EXPERIMENT_CONFIG['NUM_RUNS']} completed!")
    print("=" * 80)


if __name__ == "__main__":
    main()