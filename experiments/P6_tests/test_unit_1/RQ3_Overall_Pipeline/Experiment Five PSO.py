import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

def execute_validation_rules(x, temp, z):
    """
    类别1: 多变量协同控制系统 (94个复杂条件) - 布尔比较格式
    已删除100%覆盖率的变异分支并重新编号

    参数:
        x: 光照强度 (lux, 范围: 1-100)
        temp: 温度 (°C, 范围: 10-40)
        z: CO2浓度 (ppm, 范围: 10-60)
    """
    # 固定值设置
    y = 60  # 土壤湿度固定为60%
    humidity = 65  # 空气湿度固定为65%
    light_ideal_high = 70

    triggered = set()

    # 分支1-9: 光照与温度协同控制
    if (x > 80 and temp > 26 and (x * 0.1 + temp) > 35) != (x < 80 and temp > 26 and (x * 0.1 + temp) > 35):
        triggered.add(1)
    if (x < 30 and temp < 20 and (x * 0.1 + temp) < 22) != (x > 30 and temp < 20 and (x * 0.1 + temp) < 22):
        triggered.add(2)
    if (x < 28 and temp < 20 and (x * 0.1 + temp) < 100) != (x > 28 and temp < 20 and (x * 0.1 + temp) < 100):
        triggered.add(3)
    if ((x - 60) * 0.5 > (temp - 24) and z < 40) != ((x - 60) * 0.5 < (temp - 24) and z < 40):
        triggered.add(4)
    if ((x - 60) * 0.5 > (temp - 24) and z < 220) != ((x - 60) * 0.5 < (temp - 24) and z < 220):
        triggered.add(5)
    if (abs(x - 60) < 15 and abs(temp - 24) < 3 and abs(z - 30) < 10 and humidity > 60) != (
            abs(x - 60) > 15 and abs(temp - 24) < 3 and abs(z - 30) < 10 and humidity > 60):
        triggered.add(6)
    if ((x * temp) > 2000 and z > 35) != ((x * temp) < 2000 and z > 35):
        triggered.add(7)
    if ((x + temp) < 100 and z < 400) != ((x + temp) > 100 and z < 400):
        triggered.add(8)
    if ((x + temp) < 100 and y < 580) != ((x + temp) > 100 and y < 580):
        triggered.add(9)

    # 分支10-18: 温度控制与复杂组合
    if (x > 85 and temp > 25 and (x / temp) > 3) != (x < 85 and temp > 25 and (x / temp) > 3):
        triggered.add(10)
    if ((x - light_ideal_high) > 5 and (temp - 28) > 1 and z > 30) != (
            (x - light_ideal_high) < 5 and (temp - 28) > 1 and z > 30):
        triggered.add(11)
    if ((x - light_ideal_high) > 5 and (temp - 28) > 8) != ((x - light_ideal_high) < 5 and (temp - 28) > 8):
        triggered.add(12)
    if ((60 - x) > (24 - temp) * 2 and z < 35) != ((60 - x) < (24 - temp) * 2 and z < 35):
        triggered.add(13)
    if ((x * 0.5) + (temp * 0.3) > 30 and z < 35) != ((x * 0.5) + (temp * 0.3) < 30 and z < 35):
        triggered.add(14)
    if ((x * 0.5) + (temp * 0.3) > 30 and z < 350) != ((x * 0.5) + (temp * 0.3) < 30 and z < 350):
        triggered.add(15)
    if (temp > 25 and z > 40 and (temp + z) > 65) != (temp < 25 and z > 40 and (temp + z) > 65):
        triggered.add(16)
    if (temp < 22 and z < 25 and (temp + z) < 45) != (temp > 22 and z < 25 and (temp + z) < 45):
        triggered.add(17)
    if ((z - 30) * 0.8 > (temp - 24) and x > 75) != ((z - 30) * 0.8 < (temp - 24) and x > 75):
        triggered.add(18)

    # 分支19-27: 光照与CO2协同控制
    if (abs(z - 30) < 10 and abs(temp - 24) < 3 and abs(y - 50) < 15) != (
            abs(z - 30) > 10 and abs(temp - 24) < 3 and abs(y - 50) < 15):
        triggered.add(19)
    if ((z * temp) > 1000 and x > 80) != ((z * temp) < 1000 and x > 80):
        triggered.add(20)
    if (x < 40 and z < 25 and (x + z) < 60) != (x > 40 and z < 25 and (x + z) < 60):
        triggered.add(21)
    if ((humidity + temp) < 85 and y < 530) != ((humidity + temp) > 85 and y < 530):
        triggered.add(22)
    if (x > 80 and z > 40 and (x / z) > 2) != (x < 80 and z > 40 and (x / z) > 2):
        triggered.add(23)
    if (z > 45 and x > 75 and (z - 40) > 3) != (z < 45 and x > 75 and (z - 40) > 3):
        triggered.add(24)
    if (z < 25 and x < 40 and (30 - z) > (60 - x) * 0.5) != (z > 25 and x < 40 and (30 - z) > (60 - x) * 0.5):
        triggered.add(25)
    if ((x * 0.3) + (z * 0.4) > 35 and temp > 25) != ((x * 0.3) + (z * 0.4) < 35 and temp > 25):
        triggered.add(26)
    if (x > 80 and temp > 26 and z > 40 and (x * 0.2 + temp * 0.3 + z * 0.1) > 30) != (
            x < 80 and temp > 26 and z > 40 and (x * 0.2 + temp * 0.3 + z * 0.1) > 30):
        triggered.add(27)

    # 分支28-36: 三变量复杂控制
    if (x < 30 and temp < 20 and z < 25 and (x + temp + z) < 70) != (
            x > 30 and temp < 20 and z < 25 and (x + temp + z) < 70):
        triggered.add(28)
    if (z < 18 and x < 30 and (z * 0.2 + x * 0.1) < 75) != (z > 18 and x < 30 and (z * 0.2 + x * 0.1) < 75):
        triggered.add(29)
    if ((z - 30) * 0.5 > (x - 60) * 0.1 and temp > 26) != ((z - 30) * 0.5 < (x - 60) * 0.1 and temp > 26):
        triggered.add(30)
    if (abs(z - 30) < 10 and abs(x - 60) < 15 and abs(temp - 24) < 2) != (
            abs(z - 30) > 10 and abs(x - 60) < 15 and abs(temp - 24) < 2):
        triggered.add(31)
    if (abs(z - 30) < 10 and abs(x - 60) < 15 and abs(temp - 24) < 108) != (
            abs(z - 30) > 10 and abs(x - 60) < 15 and abs(temp - 24) < 108):
        triggered.add(32)
    if ((z * x) > 3000 and temp > 25) != ((z * x) < 3000 and temp > 25):
        triggered.add(33)
    if (x < 35 and temp < 22 and z < 28) != (x > 35 and temp < 22 and z < 28):
        triggered.add(34)
    if (x > 75 and temp > 24 and z > 35) != (x < 75 and temp > 24 and z > 35):
        triggered.add(35)
    if ((x + temp + z) > 150) != ((x + temp + z) < 150):
        triggered.add(36)

    # 分支37-46: 比值与关系控制
    if ((x * temp * z) > 50000) != ((x * temp * z) < 50000):
        triggered.add(37)
    if (abs(x - 60) + abs(temp - 24) + abs(z - 30) > 40) != (abs(x - 60) + abs(temp - 24) + abs(z - 30) < 40):
        triggered.add(38)
    if (x / (temp + 1) > 3) != (x / (temp + 1) < 3):
        triggered.add(39)
    if (z / (x + 1) > 0.8) != (z / (x + 1) < 0.8):
        triggered.add(40)
    if (temp / (z + 1) > 0.8) != (temp / (z + 1) < 0.8):
        triggered.add(41)
    if ((x - 60) * (temp - 24) > 100) != ((x - 60) * (temp - 24) < 100):
        triggered.add(42)
    if ((z - 30) * (x - 60) > 200) != ((z - 30) * (x - 60) < 200):
        triggered.add(43)
    if (temp > 28 and x > 70) != (temp < 28 and x > 70):
        triggered.add(44)
    if (temp < 18 and z < 30) != (temp > 18 and z < 30):
        triggered.add(45)
    if (x > 85 and z > 35) != (x < 85 and z > 35):
        triggered.add(46)

    # 分支47-56: 组合条件控制
    if (x < 25 and temp < 22) != (x > 25 and temp < 22):
        triggered.add(47)
    if (z > 50 and temp > 26) != (z < 50 and temp > 26):
        triggered.add(48)
    if (z < 20 and x < 35) != (z > 20 and x < 35):
        triggered.add(49)
    if (x > 75 and temp > 25 and z > 38) != (x < 75 and temp > 25 and z > 38):
        triggered.add(50)
    if (x < 35 and temp < 21 and z < 28) != (x > 35 and temp < 21 and z < 28):
        triggered.add(51)
    if (x > 0 and temp > 0 and (x / temp) > 3.5) != (x > 0 and temp > 0 and (x / temp) < 3.5):
        triggered.add(52)
    if (z > 0 and x > 0 and (z / x) > 0.7) != (z > 0 and x > 0 and (z / x) < 0.7):
        triggered.add(53)
    if ((x - temp) > 50) != ((x - temp) < 50):
        triggered.add(54)
    if ((z - temp) > 10) != ((z - temp) < 10):
        triggered.add(55)
    if ((x + temp) > 110) != ((x + temp) < 110):
        triggered.add(56)

    # 分支57-62: 复杂表达式控制
    if ((z + x) > 120) != ((z + x) < 120):
        triggered.add(57)
    if ((x * 0.4 + temp * 0.3 + z * 0.1) > 32) != ((x * 0.4 + temp * 0.3 + z * 0.1) < 32):
        triggered.add(58)
    if ((x - 60) ** 2 + (temp - 24) ** 2 > 500) != ((x - 60) ** 2 + (temp - 24) ** 2 < 500):
        triggered.add(59)
    if (x * temp * z > 60000) != (x * temp * z < 60000):
        triggered.add(60)
    if (abs(x - 60) + abs(temp - 24) + abs(z - 30) > 35) != (abs(x - 60) + abs(temp - 24) + abs(z - 30) < 35):
        triggered.add(61)
    if ((x > 70 and temp < 20) or (x < 30 and temp > 28)) != ((x < 70 and temp < 20) or (x < 30 and temp > 28)):
        triggered.add(62)

    # 分支63-68: 边界组合控制
    if ((z > 40 and x < 35) or (z < 25 and x > 75)) != ((z < 40 and x < 35) or (z < 25 and x > 75)):
        triggered.add(63)
    if (x >= 90 and temp >= 35) != (x < 90 and temp >= 35):
        triggered.add(64)
    if (x <= 10 and temp <= 15) != (x > 10 and temp <= 15):
        triggered.add(65)
    if (z >= 52 and x >= 88) != (z < 52 and x >= 88):
        triggered.add(66)
    if (z <= 18 and x <= 12) != (z > 18 and x <= 12):
        triggered.add(67)
    if (x > 0 and temp > 0 and (x / temp) >= 4) != (x > 0 and temp > 0 and (x / temp) < 4):
        triggered.add(68)

    # 分支69-72: 极端情况控制
    if (temp > 0 and x > 0 and (temp / x) >= 1) != (temp > 0 and x > 0 and (temp / x) < 1):
        triggered.add(69)
    if (x >= 90 and temp >= 35 and z >= 50) != (x < 90 and temp >= 35 and z >= 50):
        triggered.add(70)
    if (x <= 10 and temp <= 15 and z <= 20) != (x > 10 and temp <= 15 and z <= 20):
        triggered.add(71)
    if ((x > 85 and temp < 18) or (x < 15 and temp > 35)) != ((x < 85 and temp < 18) or (x < 15 and temp > 35)):
        triggered.add(72)

    # 分支73-94: 精细化控制
    if ((z > 50 and x < 15) or (z < 18 and x > 88)) != ((z < 50 and x < 15) or (z < 18 and x > 88)):
        triggered.add(73)
    if (x > 65 and temp > 27 and z > 42 and (x + temp + z) > 135) != (
            x < 65 and temp > 27 and z > 42 and (x + temp + z) > 135):
        triggered.add(74)
    if (x < 45 and temp < 23 and z < 32 and (x + temp + z) < 95) != (
            x > 45 and temp < 23 and z < 32 and (x + temp + z) < 95):
        triggered.add(75)
    if ((x / (temp + 1)) * (z / (x + 1)) > 1.5) != ((x / (temp + 1)) * (z / (x + 1)) < 1.5):
        triggered.add(76)
    if (40 <= x <= 80 and 22 <= temp <= 26 and 28 <= z <= 35) != (
            not (40 <= x <= 80) and 22 <= temp <= 26 and 28 <= z <= 35):
        triggered.add(77)
    if (abs(x - 60) * abs(temp - 24) * abs(z - 30) > 500) != (abs(x - 60) * abs(temp - 24) * abs(z - 30) < 500):
        triggered.add(78)
    if (x > 60 and temp > 24 and z > 30 and (x - 60) + (temp - 24) + (z - 30) > 20) != (
            x < 60 and temp > 24 and z > 30 and (x - 60) + (temp - 24) + (z - 30) > 20):
        triggered.add(79)
    if (x < 60 and temp < 24 and z < 30 and (60 - x) + (24 - temp) + (30 - z) > 20) != (
            x > 60 and temp < 24 and z < 30 and (60 - x) + (24 - temp) + (30 - z) > 20):
        triggered.add(80)
    if ((x * 0.3 + temp * 0.5 + z * 0.2) > 35) != ((x * 0.3 + temp * 0.5 + z * 0.2) < 35):
        triggered.add(81)
    if ((x * 0.3 + temp * 0.5 + z * 0.2) < 20) != ((x * 0.3 + temp * 0.5 + z * 0.2) > 20):
        triggered.add(82)
    if (x > 70 and z > 35 and (x - z) > 30) != (x < 70 and z > 35 and (x - z) > 30):
        triggered.add(83)
    if (z > 40 and x < 50 and (z - x) > 10) != (z < 40 and x < 50 and (z - x) > 10):
        triggered.add(84)
    if (temp > 26 and (x + z) > 120) != (temp < 26 and (x + z) > 120):
        triggered.add(85)
    if (temp < 22 and (x + z) < 80) != (temp > 22 and (x + z) < 80):
        triggered.add(86)
    if (abs(x - temp) < 10 and abs(temp - z) < 10) != (abs(x - temp) > 10 and abs(temp - z) < 10):
        triggered.add(87)
    if (max(x, temp, z) - min(x, temp, z) > 60) != (max(x, temp, z) - min(x, temp, z) < 60):
        triggered.add(88)
    if ((x > 80 or temp > 30 or z > 45) and not (x > 80 and temp > 30 and z > 45)) != (
            (x < 80 or temp > 30 or z > 45) and not (x > 80 and temp > 30 and z > 45)):
        triggered.add(89)
    if ((x < 30 or temp < 20 or z < 25) and not (x < 30 and temp < 20 and z < 25)) != (
            (x > 30 or temp < 20 or z < 25) and not (x < 30 and temp < 20 and z < 25)):
        triggered.add(90)
    if (x + temp > 120 and z < 30) != (x + temp < 120 and z < 30):
        triggered.add(91)
    if (x + z > 130 and temp < 20) != (x + z < 130 and temp < 20):
        triggered.add(92)
    if (temp + z > 70 and x < 40) != (temp + z < 70 and x < 40):
        triggered.add(93)
    if ((x - 50) ** 2 + (temp - 25) ** 2 + (z - 35) ** 2 > 1000) != (
            (x - 50) ** 2 + (temp - 25) ** 2 + (z - 35) ** 2 < 1000):
        triggered.add(94)

    return triggered

def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """"""
    generated_path = execute_validation_rules(particle[0], particle[1], particle[2])

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # : x:1-50, y:1-50, z:1-50
        self.bounds = bounds if bounds else [(1, 100), (10, 40), (10, 60)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """target pathsPSO"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(particles[i][0], particles[i][1], particles[i][2]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(gbest_particle[0], gbest_particle[1], gbest_particle[2]),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """Path PSO"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - Path ")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations")
    print(f"Path : {len(target_paths)}")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"Path {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "" if result['success'] else f"({result['best_fitness']:.3f})"
        print(f"{status} | {result['time']:.2f}s | iterations{result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f": {success_count}/{len(target_paths)} ({success_rate:.1f}%) | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs}")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations, {len(target_paths)}Path ")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"---  {run_idx}/{num_runs} ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f": {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs} runcompleted | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 1: 
    ws1 = wb.active
    ws1.title = ""
    ws1.sheet_view.showGridLines = False

    headers = ["Run", "", "", "", "Average Iterations", "(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f" {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # 2: Path 
    ws2 = wb.create_sheet(title="Path ")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Path ID", "", "", "", "Average Iterations", "Minimum Iterations", "Maximum Iterations"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"Path  {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # 3: 
    ws3 = wb.create_sheet(title="")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Path ", "", "(x,y,z)", "", "Iterations", "Path "]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"{run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # 4: target paths
    ws4 = wb.create_sheet(title="target paths")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Path ID", "target paths", ""]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"Path  {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f" : {filename}")
    print(f"{'=' * 70}")
    print(f":")
    print(f"  1.        - {len(all_results)} run")
    print(f"  2. Path        - Path ")
    print(f"  3.    -  runPath ")
    print(f"  4. target paths       - target paths")
    print(f"{'=' * 70}\n")

    return filename


def main():


    target_paths = [
        [2, 3, 4, 5, 8, 9, 15, 18, 22, 36, 37, 39, 40, 41, 42, 43, 44, 47, 52, 53, 54, 55, 56, 57, 58, 59, 60, 61, 62,
         65, 68,
         69, 72, 76, 78, 81, 82, 88, 89, 92, 94],
        [1, 4, 5, 7, 8, 9, 11, 12, 15, 22, 26, 29, 30, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 49, 52, 53, 54,
         55, 56,
         57, 58, 59, 60, 61, 63, 64, 67, 68, 69, 73, 76, 78, 81, 82, 84, 86, 88, 90, 93, 94],
        [1, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52, 53,
         54,
         55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
        [4, 5, 8, 9, 15, 22, 36, 37, 39, 40, 41, 42, 44, 47, 52, 53, 54, 55, 56, 57, 58, 59, 60, 61, 65, 68, 69, 72, 76,
         78,
         81, 82, 88, 89, 92, 94],
        [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 23, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50,
         52, 53,
         54, 55, 56, 57, 58, 59, 60, 61, 63, 64, 67, 68, 69, 70, 73, 76, 78, 79, 81, 82, 84, 85, 87, 88, 90, 93, 94],
        [1, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52,
         53, 54,
         55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
        [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52,
         53, 54,
         55, 56, 57, 58, 59, 60, 61, 63, 64, 67, 68, 69, 70, 73, 76, 78, 79, 81, 82, 84, 85, 87, 88, 90, 93, 94],
        [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52,
         53, 54,
         55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
        [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52,
         53, 54,
         55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
        [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52,
         53, 54,
         55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
        [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52,
         53, 54,
         55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
        [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52,
         53, 54,
         55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
        [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52,
         53, 54,
         55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
        [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52,
         53, 54,
         55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
        [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52,
         53, 54,
         55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
        [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52,
         53, 54,
         55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
        [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52,
         53, 54,
         55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
        [1, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52,
         53, 54,
         55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
        [1, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52,
         53, 54,
         55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
        [1, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52,
         53, 54,
         55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
        [1, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52,
         53, 54,
         55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
        [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52,
         53, 54,
         55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
        [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52,
         53, 54,
         55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
        [1, 4, 5, 7, 8, 9, 11, 12, 15, 22, 26, 29, 30, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 49, 52, 53, 54,
         55, 56,
         57, 58, 59, 60, 61, 63, 64, 67, 68, 69, 73, 76, 78, 81, 82, 84, 86, 88, 90, 93, 94],
        [1, 4, 5, 7, 8, 9, 11, 12, 15, 22, 26, 29, 30, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 49, 52, 53, 54,
         55, 56,
         57, 58, 59, 60, 61, 63, 64, 67, 68, 69, 73, 76, 78, 81, 82, 84, 86, 88, 90, 93, 94]
    ]

    print("=" * 70)
    print("baseline PSO")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("Program completed")


if __name__ == "__main__":
    main()