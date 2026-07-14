import os
import random
from collections import deque
import numpy as np
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# --- 全局状态范围配置 ---
STATE_MIN_X, STATE_MAX_X = 1, 100
STATE_MIN_Y, STATE_MAX_Y = 10, 40
STATE_MIN_Z, STATE_MAX_Z = 10, 60


def generate_input():
    return [
        random.randint(STATE_MIN_X, STATE_MAX_X),
        random.randint(STATE_MIN_Y, STATE_MAX_Y),
        random.randint(STATE_MIN_Z, STATE_MAX_Z)
    ]


def execute_Tr(x, temp, z):
    """
    类别1: 多变量协同控制系统 (94个复杂条件) - 布尔比较格式
    已删除100%覆盖率的变异分支并重新编号

    参数:
        x: 光照强度 (lux, 范围: 1-100)
        temp: 温度 (°C, 范围: 10-40)
        z: CO2浓度 (ppm, 范围: 10-60)
    """
    # 固定值设置
    y = 60  # 土壤湿度固定为60%
    humidity = 65  # 空气湿度固定为65%
    light_ideal_high = 70

    triggered = set()

    # 分支1-9: 光照与温度协同控制
    if (x > 80 and temp > 26 and (x * 0.1 + temp) > 35) != (x < 80 and temp > 26 and (x * 0.1 + temp) > 35):
        triggered.add(1)
    if (x < 30 and temp < 20 and (x * 0.1 + temp) < 22) != (x > 30 and temp < 20 and (x * 0.1 + temp) < 22):
        triggered.add(2)
    if (x < 28 and temp < 20 and (x * 0.1 + temp) < 100) != (x > 28 and temp < 20 and (x * 0.1 + temp) < 100):
        triggered.add(3)
    if ((x - 60) * 0.5 > (temp - 24) and z < 40) != ((x - 60) * 0.5 < (temp - 24) and z < 40):
        triggered.add(4)
    if ((x - 60) * 0.5 > (temp - 24) and z < 220) != ((x - 60) * 0.5 < (temp - 24) and z < 220):
        triggered.add(5)
    if (abs(x - 60) < 15 and abs(temp - 24) < 3 and abs(z - 30) < 10 and humidity > 60) != (
            abs(x - 60) > 15 and abs(temp - 24) < 3 and abs(z - 30) < 10 and humidity > 60):
        triggered.add(6)
    if ((x * temp) > 2000 and z > 35) != ((x * temp) < 2000 and z > 35):
        triggered.add(7)
    if ((x + temp) < 100 and z < 400) != ((x + temp) > 100 and z < 400):
        triggered.add(8)
    if ((x + temp) < 100 and y < 580) != ((x + temp) > 100 and y < 580):
        triggered.add(9)

    # 分支10-18: 温度控制与复杂组合
    if (x > 85 and temp > 25 and (x / temp) > 3) != (x < 85 and temp > 25 and (x / temp) > 3):
        triggered.add(10)
    if ((x - light_ideal_high) > 5 and (temp - 28) > 1 and z > 30) != (
            (x - light_ideal_high) < 5 and (temp - 28) > 1 and z > 30):
        triggered.add(11)
    if ((x - light_ideal_high) > 5 and (temp - 28) > 8) != ((x - light_ideal_high) < 5 and (temp - 28) > 8):
        triggered.add(12)
    if ((60 - x) > (24 - temp) * 2 and z < 35) != ((60 - x) < (24 - temp) * 2 and z < 35):
        triggered.add(13)
    if ((x * 0.5) + (temp * 0.3) > 30 and z < 35) != ((x * 0.5) + (temp * 0.3) < 30 and z < 35):
        triggered.add(14)
    if ((x * 0.5) + (temp * 0.3) > 30 and z < 350) != ((x * 0.5) + (temp * 0.3) < 30 and z < 350):
        triggered.add(15)
    if (temp > 25 and z > 40 and (temp + z) > 65) != (temp < 25 and z > 40 and (temp + z) > 65):
        triggered.add(16)
    if (temp < 22 and z < 25 and (temp + z) < 45) != (temp > 22 and z < 25 and (temp + z) < 45):
        triggered.add(17)
    if ((z - 30) * 0.8 > (temp - 24) and x > 75) != ((z - 30) * 0.8 < (temp - 24) and x > 75):
        triggered.add(18)

    # 分支19-27: 光照与CO2协同控制
    if (abs(z - 30) < 10 and abs(temp - 24) < 3 and abs(y - 50) < 15) != (
            abs(z - 30) > 10 and abs(temp - 24) < 3 and abs(y - 50) < 15):
        triggered.add(19)
    if ((z * temp) > 1000 and x > 80) != ((z * temp) < 1000 and x > 80):
        triggered.add(20)
    if (x < 40 and z < 25 and (x + z) < 60) != (x > 40 and z < 25 and (x + z) < 60):
        triggered.add(21)
    if ((humidity + temp) < 85 and y < 530) != ((humidity + temp) > 85 and y < 530):
        triggered.add(22)
    if (x > 80 and z > 40 and (x / z) > 2) != (x < 80 and z > 40 and (x / z) > 2):
        triggered.add(23)
    if (z > 45 and x > 75 and (z - 40) > 3) != (z < 45 and x > 75 and (z - 40) > 3):
        triggered.add(24)
    if (z < 25 and x < 40 and (30 - z) > (60 - x) * 0.5) != (z > 25 and x < 40 and (30 - z) > (60 - x) * 0.5):
        triggered.add(25)
    if ((x * 0.3) + (z * 0.4) > 35 and temp > 25) != ((x * 0.3) + (z * 0.4) < 35 and temp > 25):
        triggered.add(26)
    if (x > 80 and temp > 26 and z > 40 and (x * 0.2 + temp * 0.3 + z * 0.1) > 30) != (
            x < 80 and temp > 26 and z > 40 and (x * 0.2 + temp * 0.3 + z * 0.1) > 30):
        triggered.add(27)

    # 分支28-36: 三变量复杂控制
    if (x < 30 and temp < 20 and z < 25 and (x + temp + z) < 70) != (
            x > 30 and temp < 20 and z < 25 and (x + temp + z) < 70):
        triggered.add(28)
    if (z < 18 and x < 30 and (z * 0.2 + x * 0.1) < 75) != (z > 18 and x < 30 and (z * 0.2 + x * 0.1) < 75):
        triggered.add(29)
    if ((z - 30) * 0.5 > (x - 60) * 0.1 and temp > 26) != ((z - 30) * 0.5 < (x - 60) * 0.1 and temp > 26):
        triggered.add(30)
    if (abs(z - 30) < 10 and abs(x - 60) < 15 and abs(temp - 24) < 2) != (
            abs(z - 30) > 10 and abs(x - 60) < 15 and abs(temp - 24) < 2):
        triggered.add(31)
    if (abs(z - 30) < 10 and abs(x - 60) < 15 and abs(temp - 24) < 108) != (
            abs(z - 30) > 10 and abs(x - 60) < 15 and abs(temp - 24) < 108):
        triggered.add(32)
    if ((z * x) > 3000 and temp > 25) != ((z * x) < 3000 and temp > 25):
        triggered.add(33)
    if (x < 35 and temp < 22 and z < 28) != (x > 35 and temp < 22 and z < 28):
        triggered.add(34)
    if (x > 75 and temp > 24 and z > 35) != (x < 75 and temp > 24 and z > 35):
        triggered.add(35)
    if ((x + temp + z) > 150) != ((x + temp + z) < 150):
        triggered.add(36)

    # 分支37-46: 比值与关系控制
    if ((x * temp * z) > 50000) != ((x * temp * z) < 50000):
        triggered.add(37)
    if (abs(x - 60) + abs(temp - 24) + abs(z - 30) > 40) != (abs(x - 60) + abs(temp - 24) + abs(z - 30) < 40):
        triggered.add(38)
    if (x / (temp + 1) > 3) != (x / (temp + 1) < 3):
        triggered.add(39)
    if (z / (x + 1) > 0.8) != (z / (x + 1) < 0.8):
        triggered.add(40)
    if (temp / (z + 1) > 0.8) != (temp / (z + 1) < 0.8):
        triggered.add(41)
    if ((x - 60) * (temp - 24) > 100) != ((x - 60) * (temp - 24) < 100):
        triggered.add(42)
    if ((z - 30) * (x - 60) > 200) != ((z - 30) * (x - 60) < 200):
        triggered.add(43)
    if (temp > 28 and x > 70) != (temp < 28 and x > 70):
        triggered.add(44)
    if (temp < 18 and z < 30) != (temp > 18 and z < 30):
        triggered.add(45)
    if (x > 85 and z > 35) != (x < 85 and z > 35):
        triggered.add(46)

    # 分支47-56: 组合条件控制
    if (x < 25 and temp < 22) != (x > 25 and temp < 22):
        triggered.add(47)
    if (z > 50 and temp > 26) != (z < 50 and temp > 26):
        triggered.add(48)
    if (z < 20 and x < 35) != (z > 20 and x < 35):
        triggered.add(49)
    if (x > 75 and temp > 25 and z > 38) != (x < 75 and temp > 25 and z > 38):
        triggered.add(50)
    if (x < 35 and temp < 21 and z < 28) != (x > 35 and temp < 21 and z < 28):
        triggered.add(51)
    if (x > 0 and temp > 0 and (x / temp) > 3.5) != (x > 0 and temp > 0 and (x / temp) < 3.5):
        triggered.add(52)
    if (z > 0 and x > 0 and (z / x) > 0.7) != (z > 0 and x > 0 and (z / x) < 0.7):
        triggered.add(53)
    if ((x - temp) > 50) != ((x - temp) < 50):
        triggered.add(54)
    if ((z - temp) > 10) != ((z - temp) < 10):
        triggered.add(55)
    if ((x + temp) > 110) != ((x + temp) < 110):
        triggered.add(56)

    # 分支57-62: 复杂表达式控制
    if ((z + x) > 120) != ((z + x) < 120):
        triggered.add(57)
    if ((x * 0.4 + temp * 0.3 + z * 0.1) > 32) != ((x * 0.4 + temp * 0.3 + z * 0.1) < 32):
        triggered.add(58)
    if ((x - 60) ** 2 + (temp - 24) ** 2 > 500) != ((x - 60) ** 2 + (temp - 24) ** 2 < 500):
        triggered.add(59)
    if (x * temp * z > 60000) != (x * temp * z < 60000):
        triggered.add(60)
    if (abs(x - 60) + abs(temp - 24) + abs(z - 30) > 35) != (abs(x - 60) + abs(temp - 24) + abs(z - 30) < 35):
        triggered.add(61)
    if ((x > 70 and temp < 20) or (x < 30 and temp > 28)) != ((x < 70 and temp < 20) or (x < 30 and temp > 28)):
        triggered.add(62)

    # 分支63-68: 边界组合控制
    if ((z > 40 and x < 35) or (z < 25 and x > 75)) != ((z < 40 and x < 35) or (z < 25 and x > 75)):
        triggered.add(63)
    if (x >= 90 and temp >= 35) != (x < 90 and temp >= 35):
        triggered.add(64)
    if (x <= 10 and temp <= 15) != (x > 10 and temp <= 15):
        triggered.add(65)
    if (z >= 52 and x >= 88) != (z < 52 and x >= 88):
        triggered.add(66)
    if (z <= 18 and x <= 12) != (z > 18 and x <= 12):
        triggered.add(67)
    if (x > 0 and temp > 0 and (x / temp) >= 4) != (x > 0 and temp > 0 and (x / temp) < 4):
        triggered.add(68)

    # 分支69-72: 极端情况控制
    if (temp > 0 and x > 0 and (temp / x) >= 1) != (temp > 0 and x > 0 and (temp / x) < 1):
        triggered.add(69)
    if (x >= 90 and temp >= 35 and z >= 50) != (x < 90 and temp >= 35 and z >= 50):
        triggered.add(70)
    if (x <= 10 and temp <= 15 and z <= 20) != (x > 10 and temp <= 15 and z <= 20):
        triggered.add(71)
    if ((x > 85 and temp < 18) or (x < 15 and temp > 35)) != ((x < 85 and temp < 18) or (x < 15 and temp > 35)):
        triggered.add(72)

    # 分支73-94: 精细化控制
    if ((z > 50 and x < 15) or (z < 18 and x > 88)) != ((z < 50 and x < 15) or (z < 18 and x > 88)):
        triggered.add(73)
    if (x > 65 and temp > 27 and z > 42 and (x + temp + z) > 135) != (
            x < 65 and temp > 27 and z > 42 and (x + temp + z) > 135):
        triggered.add(74)
    if (x < 45 and temp < 23 and z < 32 and (x + temp + z) < 95) != (
            x > 45 and temp < 23 and z < 32 and (x + temp + z) < 95):
        triggered.add(75)
    if ((x / (temp + 1)) * (z / (x + 1)) > 1.5) != ((x / (temp + 1)) * (z / (x + 1)) < 1.5):
        triggered.add(76)
    if (40 <= x <= 80 and 22 <= temp <= 26 and 28 <= z <= 35) != (
            not (40 <= x <= 80) and 22 <= temp <= 26 and 28 <= z <= 35):
        triggered.add(77)
    if (abs(x - 60) * abs(temp - 24) * abs(z - 30) > 500) != (abs(x - 60) * abs(temp - 24) * abs(z - 30) < 500):
        triggered.add(78)
    if (x > 60 and temp > 24 and z > 30 and (x - 60) + (temp - 24) + (z - 30) > 20) != (
            x < 60 and temp > 24 and z > 30 and (x - 60) + (temp - 24) + (z - 30) > 20):
        triggered.add(79)
    if (x < 60 and temp < 24 and z < 30 and (60 - x) + (24 - temp) + (30 - z) > 20) != (
            x > 60 and temp < 24 and z < 30 and (60 - x) + (24 - temp) + (30 - z) > 20):
        triggered.add(80)
    if ((x * 0.3 + temp * 0.5 + z * 0.2) > 35) != ((x * 0.3 + temp * 0.5 + z * 0.2) < 35):
        triggered.add(81)
    if ((x * 0.3 + temp * 0.5 + z * 0.2) < 20) != ((x * 0.3 + temp * 0.5 + z * 0.2) > 20):
        triggered.add(82)
    if (x > 70 and z > 35 and (x - z) > 30) != (x < 70 and z > 35 and (x - z) > 30):
        triggered.add(83)
    if (z > 40 and x < 50 and (z - x) > 10) != (z < 40 and x < 50 and (z - x) > 10):
        triggered.add(84)
    if (temp > 26 and (x + z) > 120) != (temp < 26 and (x + z) > 120):
        triggered.add(85)
    if (temp < 22 and (x + z) < 80) != (temp > 22 and (x + z) < 80):
        triggered.add(86)
    if (abs(x - temp) < 10 and abs(temp - z) < 10) != (abs(x - temp) > 10 and abs(temp - z) < 10):
        triggered.add(87)
    if (max(x, temp, z) - min(x, temp, z) > 60) != (max(x, temp, z) - min(x, temp, z) < 60):
        triggered.add(88)
    if ((x > 80 or temp > 30 or z > 45) and not (x > 80 and temp > 30 and z > 45)) != (
            (x < 80 or temp > 30 or z > 45) and not (x > 80 and temp > 30 and z > 45)):
        triggered.add(89)
    if ((x < 30 or temp < 20 or z < 25) and not (x < 30 and temp < 20 and z < 25)) != (
            (x > 30 or temp < 20 or z < 25) and not (x < 30 and temp < 20 and z < 25)):
        triggered.add(90)
    if (x + temp > 120 and z < 30) != (x + temp < 120 and z < 30):
        triggered.add(91)
    if (x + z > 130 and temp < 20) != (x + z < 130 and temp < 20):
        triggered.add(92)
    if (temp + z > 70 and x < 40) != (temp + z < 70 and x < 40):
        triggered.add(93)
    if ((x - 50) ** 2 + (temp - 25) ** 2 + (z - 35) ** 2 > 1000) != (
            (x - 50) ** 2 + (temp - 25) ** 2 + (z - 35) ** 2 < 1000):
        triggered.add(94)

    return triggered


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0


# Fixed: Proper Python list syntax
TARGET_PATHS = [
    [2, 3, 4, 5, 8, 9, 15, 18, 22, 36, 37, 39, 40, 41, 42, 43, 44, 47, 52, 53, 54, 55, 56, 57, 58, 59, 60, 61, 62, 65, 68,
     69, 72, 76, 78, 81, 82, 88, 89, 92, 94],
    [1, 4, 5, 7, 8, 9, 11, 12, 15, 22, 26, 29, 30, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 49, 52, 53, 54, 55, 56,
     57, 58, 59, 60, 61, 63, 64, 67, 68, 69, 73, 76, 78, 81, 82, 84, 86, 88, 90, 93, 94],
    [1, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52, 53, 54,
     55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
    [4, 5, 8, 9, 15, 22, 36, 37, 39, 40, 41, 42, 44, 47, 52, 53, 54, 55, 56, 57, 58, 59, 60, 61, 65, 68, 69, 72, 76, 78,
     81, 82, 88, 89, 92, 94],
    [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 23, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52, 53,
     54, 55, 56, 57, 58, 59, 60, 61, 63, 64, 67, 68, 69, 70, 73, 76, 78, 79, 81, 82, 84, 85, 87, 88, 90, 93, 94],
    [1, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52, 53, 54,
     55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
    [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52, 53, 54,
     55, 56, 57, 58, 59, 60, 61, 63, 64, 67, 68, 69, 70, 73, 76, 78, 79, 81, 82, 84, 85, 87, 88, 90, 93, 94],
    [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52, 53, 54,
     55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
    [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52, 53, 54,
     55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
    [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52, 53, 54,
     55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
    [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52, 53, 54,
     55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
    [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52, 53, 54,
     55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
    [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52, 53, 54,
     55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
    [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52, 53, 54,
     55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
    [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52, 53, 54,
     55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
    [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52, 53, 54,
     55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
    [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52, 53, 54,
     55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
    [1, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52, 53, 54,
     55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
    [1, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52, 53, 54,
     55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
    [1, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52, 53, 54,
     55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
    [1, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52, 53, 54,
     55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
    [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52, 53, 54,
     55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
    [2, 4, 5, 7, 8, 9, 11, 12, 15, 16, 22, 26, 27, 30, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 50, 52, 53, 54,
     55, 56, 57, 58, 59, 60, 61, 64, 68, 69, 70, 76, 78, 79, 81, 82, 85, 87, 88, 90, 94],
    [1, 4, 5, 7, 8, 9, 11, 12, 15, 22, 26, 29, 30, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 49, 52, 53, 54, 55, 56,
     57, 58, 59, 60, 61, 63, 64, 67, 68, 69, 73, 76, 78, 81, 82, 84, 86, 88, 90, 93, 94],
    [1, 4, 5, 7, 8, 9, 11, 12, 15, 22, 26, 29, 30, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 46, 48, 49, 52, 53, 54, 55, 56,
     57, 58, 59, 60, 61, 63, 64, 67, 68, 69, 73, 76, 78, 81, 82, 84, 86, 88, 90, 93, 94]
]


class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 2000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(TARGET_PATHS)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }


def compute_robustness(state, path):
    base = execute_Tr(state[0], state[1], state[2])
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                    np.clip(state[0] + dx, STATE_MIN_X, STATE_MAX_X),
                    np.clip(state[1] + dy, STATE_MIN_Y, STATE_MAX_Y),
                    np.clip(state[2] + dz, STATE_MIN_Z, STATE_MAX_Z)
                ])
                n_trig = execute_Tr(neighbor[0], neighbor[1], neighbor[2])
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = TARGET_PATHS[target_path_idx]
    samples = []
    attempts = 0

    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(STATE_MIN_X, STATE_MAX_X),
            random.randint(STATE_MIN_Y, STATE_MAX_Y),
            random.randint(STATE_MIN_Z, STATE_MAX_Z)
        ])
        triggered = execute_Tr(state[0], state[1], state[2])

        if not triggered:
            continue

        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)

        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }

        samples.append(sample_data)

    return samples


def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample[
                'robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)

        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]

    return selected_samples


def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = TARGET_PATHS[target_path_idx]
        samples = []
        attempts = 0

        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(STATE_MIN_X, STATE_MAX_X),
                random.randint(STATE_MIN_Y, STATE_MAX_Y),
                random.randint(STATE_MIN_Z, STATE_MAX_Z)
            ])
            triggered = execute_Tr(state[0], state[1], state[2])

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)

            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)

        random.shuffle(samples)
        return samples[:config.top_k_samples]

    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")

        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)


def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}

    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates

    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}

        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config,
                                                         shared_candidates[path_idx])

            strategy_results[path_idx] = samples

        results[strategy_name] = strategy_results

    return results


def analyze_fitness_values(results, config):
    analysis_results = {}

    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []

        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])

        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }

        all_scores = []

        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])

        scores_array = np.array(all_scores)

        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)

        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })

        analysis_results[strategy_name] = analysis

    return analysis_results


def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []

    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)

    df = pd.DataFrame(df_data)
    return df, analysis_results


def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []

    print(f"Starting {num_runs} experiments...")

    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")

        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)

        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)

        print(f"Experiment {run_idx + 1} completed.")

    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df


def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')

    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width

    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []

    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]

        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)

    stats_df = pd.DataFrame(stats_data)

    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Results saved to: {output_path}")


def main():
    results_df = run_multiple_experiments(num_runs=20)

    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")

    save_results_to_excel(results_df, output_path)

    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)

    return results_df, output_path


if __name__ == "__main__":
    results, output_file = main()