import os
import sys
import random
import time
from collections import deque
from datetime import datetime
from typing import Any, Dict, List, Optional, Set, Tuple

import numpy as np
import torch
import torch.nn as nn
import torch.optim as optim
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ========================================
# ========== experiment configuration parameters ==========
# ========================================
NUM_RUNS = 20
def safe_divide(a, b):
    """安全除法，避免除零错误"""
    if b == 0:
        return 0
    return a / b
# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === three-dimensional range settings ===
# Keep the current DQN state range used by the second script. To use a 0-500 range, modify this section only.
LIGHT_MIN = 10
LIGHT_MAX = 80
MOISTURE_MIN = 800
MOISTURE_MAX = 1500
TEMP_MIN = 1
TEMP_MAX = 40
BOUNDS = {
    "light": (LIGHT_MIN, LIGHT_MAX),
    "moisture": (MOISTURE_MIN, MOISTURE_MAX),
    "temp": (TEMP_MIN, TEMP_MAX),
}

# === standard PSO parameters ===
PSO_W = 0.7
PSO_C1 = 1.5
PSO_C2 = 1.5
PSO_VMAX_RATIO = 0.2


def normalize_state(state):
    """Normalize the state to the [0, 1] range."""
    return np.array([
        (state[0] - BOUNDS["light"][0]) / (BOUNDS["light"][1] - BOUNDS["light"][0]),
        (state[1] - BOUNDS["moisture"][0]) / (BOUNDS["moisture"][1] - BOUNDS["moisture"][0]),
        (state[2] - BOUNDS["temp"][0]) / (BOUNDS["temp"][1] - BOUNDS["temp"][0]),
    ], dtype=np.float32)


def denormalize_state(normalized_state):
    """Restore the normalized state to the original range."""
    return np.array([
        normalized_state[0] * (BOUNDS["light"][1] - BOUNDS["light"][0]) + BOUNDS["light"][0],
        normalized_state[1] * (BOUNDS["moisture"][1] - BOUNDS["moisture"][0]) + BOUNDS["moisture"][0],
        normalized_state[2] * (BOUNDS["temp"][1] - BOUNDS["temp"][0]) + BOUNDS["temp"][0],
    ], dtype=float)


def generate_random_state():
    """Generate a random state within each dimension range."""
    light = np.random.randint(BOUNDS["light"][0], BOUNDS["light"][1] + 1)
    moisture = np.random.randint(BOUNDS["moisture"][0], BOUNDS["moisture"][1] + 1)
    temp = np.random.randint(BOUNDS["temp"][0], BOUNDS["temp"][1] + 1)
    return np.array([light, moisture, temp], dtype=float)


def clip_state(state):
    """Clip the state to each dimension boundary."""
    return np.array([
        np.clip(state[0], BOUNDS["light"][0], BOUNDS["light"][1]),
        np.clip(state[1], BOUNDS["moisture"][0], BOUNDS["moisture"][1]),
        np.clip(state[2], BOUNDS["temp"][0], BOUNDS["temp"][1]),
    ], dtype=float)


def is_state_valid(state):
    """Check whether the state is within all valid dimension ranges."""
    return (
        BOUNDS["light"][0] <= state[0] <= BOUNDS["light"][1]
        and BOUNDS["moisture"][0] <= state[1] <= BOUNDS["moisture"][1]
        and BOUNDS["temp"][0] <= state[2] <= BOUNDS["temp"][1]
    )


def execute_Tr(moisture, co2, temp):
    """检测低CO2极端条件下的分支触发情况"""
    triggered = set()

    # Fixed: proper indentation and using triggered.add()
    if (co2 < 1150) != (co2 < 1000):
        triggered.add(1)
    if (co2 < 1150) != (co2 < 950):
        triggered.add(2)
    if (co2 < 1150) != (co2 < 1400):
        triggered.add(3)

    # 分支4-11: 调整到中间值区域
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1050 and moisture > 45 and temp > 20):
        triggered.add(4)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1100 and moisture > 45 and temp > 20):
        triggered.add(5)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 or moisture > 45 and temp > 20):
        triggered.add(6)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 35 and temp > 20):
        triggered.add(7)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 30 and temp > 20):
        triggered.add(8)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture < 45 and temp > 20):
        triggered.add(9)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 45 and temp > 5):
        triggered.add(10)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 45 and temp < 20):
        triggered.add(11)

    # 分支12-21: 调整到中间值
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 970 and moisture > 50) or (co2 < 1100 and temp < 22)):
        triggered.add(12)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1000 and moisture > 50) or (co2 < 1100 and temp < 22)):
        triggered.add(13)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 or moisture > 50) or (co2 < 1100 and temp < 22)):
        triggered.add(14)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 40) or (co2 < 1100 and temp < 22)):
        triggered.add(15)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 35) or (co2 < 1100 and temp < 22)):
        triggered.add(16)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 50) and (co2 < 1100 and temp < 22)):
        triggered.add(17)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 50) or (co2 < 960 and temp < 22)):
        triggered.add(18)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 18)):
        triggered.add(19)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 50) or (co2 < 1100 and temp > 22)):
        triggered.add(20)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 50) or (co2 > 1100 and temp < 22)):
        triggered.add(21)

    # 分支22-25: 调整阈值到中间 - using safe_divide
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp < 22) != (safe_divide(moisture, co2 - 700) > 0.04 and temp < 22):
        triggered.add(22)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp < 22) != (safe_divide(moisture, co2 - 700) > 0.03 and temp < 22):
        triggered.add(23)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp < 22) != (safe_divide(moisture, co2 - 700) > 0.06 or temp < 22):
        triggered.add(24)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp < 22) != (moisture + safe_divide(100, co2 - 700) > 0.06 and temp < 22):
        triggered.add(25)

    # 分支26-36: 调整到中间值
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1000 and moisture > 45 and temp > 20):
        triggered.add(26)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1070 and moisture > 45 and temp > 20):
        triggered.add(27)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 > 1150 and moisture > 45 and temp > 20):
        triggered.add(28)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture < 45 and temp > 20):
        triggered.add(29)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 25 and temp > 20):
        triggered.add(30)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 30 and temp > 20):
        triggered.add(31)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 45 and temp < 20):
        triggered.add(32)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 45 and temp > 15):
        triggered.add(33)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 45 and temp > 12):
        triggered.add(34)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 45 or temp > 20):
        triggered.add(35)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 or moisture > 45 and temp > 20):
        triggered.add(36)

    # 分支37-44: 调整到中间值
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 970 and moisture < 40 and temp < 22):
        triggered.add(37)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 1020 and moisture < 40 and temp < 22):
        triggered.add(38)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 > 1150 and moisture < 40 and temp < 22):
        triggered.add(39)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 1150 or moisture < 40 and temp < 22):
        triggered.add(40)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 1150 and moisture < 50 and temp < 22):
        triggered.add(41)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 1150 and moisture > 40 and temp < 22):
        triggered.add(42)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 1150 and moisture < 40 and temp < 27):
        triggered.add(43)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 1150 and moisture < 40 and temp < 32):
        triggered.add(44)

    # 分支45-55: 扩大范围
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1050 and 35 < moisture < 60 and 15 < temp < 28):
        triggered.add(45)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1120 and 35 < moisture < 60 and 15 < temp < 28):
        triggered.add(46)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1070 and 35 < moisture < 60 and 15 < temp < 28):
        triggered.add(47)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 > 1150 and 35 < moisture < 60 and 15 < temp < 28):
        triggered.add(48)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 25 < moisture < 60 and 15 < temp < 28):
        triggered.add(49)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 20 < moisture < 60 and 15 < temp < 28):
        triggered.add(50)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 35 < moisture < 52 and 15 < temp < 28):
        triggered.add(51)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 35 < moisture < 60 or 15 < temp < 28):
        triggered.add(52)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 35 < moisture < 60 and 10 < temp < 28):
        triggered.add(53)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 35 < moisture < 60 and 15 < temp < 20):
        triggered.add(54)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 35 < moisture < 60 and 15 < temp < 33):
        triggered.add(55)

    # 分支56-65: 调整到中间值和扩大范围
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1020 and moisture < 40 and 15 < temp < 25):
        triggered.add(56)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1070 and moisture < 40 and 15 < temp < 25):
        triggered.add(57)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 > 1150 and moisture < 40 and 15 < temp < 25):
        triggered.add(58)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 or moisture < 40 and 15 < temp < 25):
        triggered.add(59)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 and moisture < 50 and 15 < temp < 25):
        triggered.add(60)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 and moisture < 55 and 15 < temp < 25):
        triggered.add(61)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 and moisture < 40 and 10 < temp < 25):
        triggered.add(62)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 and moisture < 40 and 5 < temp < 25):
        triggered.add(63)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 and moisture < 40 and 15 < temp < 30):
        triggered.add(64)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 and moisture < 40 and 15 < temp < 30):
        triggered.add(65)

    # 分支66-74: 调整比例和温度阈值
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 850) > 0.06 and temp > 20):
        triggered.add(66)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 900) > 0.06 and temp > 20):
        triggered.add(67)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 + 200) > 0.06 and temp > 20):
        triggered.add(68)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 700) > 0.04 and temp > 20):
        triggered.add(69)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 700) > 0.03 and temp > 20):
        triggered.add(70)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 700) > 0.06 or temp > 20):
        triggered.add(71)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 700) > 0.06 and temp > 15):
        triggered.add(72)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 700) > 0.06 and temp > 12):
        triggered.add(73)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 700) > 0.06 and temp < 20):
        triggered.add(74)

    if (co2 + moisture > 1150 and temp > 18) != (co2 + moisture > 1100 and temp > 18):
        triggered.add(75)
    if (co2 + moisture > 1150 and temp > 18) != (co2 + moisture > 1050 and temp > 18):
        triggered.add(76)
    if (co2 + moisture > 1150 and temp > 18) != (co2 + moisture > 1150 and temp > 15):
        triggered.add(77)
    if (co2 + moisture > 1150 and temp > 18) != (co2 + moisture > 1150 and temp > 12):
        triggered.add(78)
    if (co2 + moisture > 1150 and temp > 18) != (co2 + moisture > 1150 and temp < 18):
        triggered.add(79)

    if (co2 < 1150 and abs(moisture - 45) > 15 and abs(temp - 20) > 7) != (
            co2 < 1000 and abs(moisture - 45) > 15 and abs(temp - 20) > 7):
        triggered.add(80)
    if (co2 < 1150 and abs(moisture - 45) > 15 and abs(temp - 20) > 7) != (
            co2 < 1050 and abs(moisture - 45) > 15 and abs(temp - 20) > 7):
        triggered.add(81)
    if (co2 < 1150 and abs(moisture - 45) > 15 and abs(temp - 20) > 7) != (
            co2 < 1150 and abs(moisture - 42) > 15 and abs(temp - 20) > 7):
        triggered.add(82)
    if (co2 < 1150 and abs(moisture - 45) > 15 and abs(temp - 20) > 7) != (
            co2 < 1150 and abs(moisture - 45) > 12 and abs(temp - 20) > 7):
        triggered.add(83)

    # 分支84-87: 简化条件到中间值
    if (co2 < 1150 and moisture < 40) != (co2 < 1000 and moisture < 40):
        triggered.add(84)
    if (co2 < 1150 and moisture < 40) != (co2 < 1050 and moisture < 40):
        triggered.add(85)
    if (co2 < 1150 and moisture < 40) != (co2 < 1150 and moisture < 50):
        triggered.add(86)
    if (co2 < 1150 and moisture < 40) != (co2 < 1150 and moisture < 55):
        triggered.add(87)

    # 分支88-92: 调整到中间范围
    if (co2 < 1150 and 12 < temp < 25 and moisture > 45) != (co2 < 1030 and 12 < temp < 25 and moisture > 45):
        triggered.add(88)
    if (co2 < 1150 and 12 < temp < 25 and moisture > 45) != (co2 < 1080 and 12 < temp < 25 and moisture > 45):
        triggered.add(89)
    if (co2 < 1150 and 12 < temp < 25 and moisture > 45) != (co2 < 1150 and 8 < temp < 25 and moisture > 45):
        triggered.add(90)
    if (co2 < 1150 and 12 < temp < 25 and moisture > 45) != (co2 < 1150 and 12 < temp < 30 and moisture > 45):
        triggered.add(91)
    if (co2 < 1150 and 12 < temp < 25 and moisture > 45) != (co2 < 1150 and 12 < temp < 25 and moisture > 35):
        triggered.add(92)

    # 分支93-96: 调整到中间值
    if (moisture > 45 and temp > 20) != (moisture > 35 and temp > 20):
        triggered.add(93)
    if (moisture > 45 and temp > 20) != (moisture > 32 and temp > 20):
        triggered.add(94)
    if (moisture > 45 and temp > 20) != (moisture > 45 and temp > 15):
        triggered.add(95)
    if (moisture > 45 and temp > 20) != (moisture > 45 and temp < 20):
        triggered.add(96)

    # 分支97-102: 扩大范围
    if (1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 25) != (
            950 < co2 < 1250 and 35 < moisture < 55 and 15 < temp < 25):
        triggered.add(97)
    if (1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 25) != (
            900 < co2 < 1300 and 35 < moisture < 55 and 15 < temp < 25):
        triggered.add(98)
    if (1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 25) != (
            1000 < co2 < 1200 and 28 < moisture < 55 and 15 < temp < 25):
        triggered.add(99)
    if (1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 25) != (
            1000 < co2 < 1200 and 35 < moisture < 62 and 15 < temp < 25):
        triggered.add(100)
    if (1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 25) != (
            1000 < co2 < 1200 and 35 < moisture < 55 and 12 < temp < 25):
        triggered.add(101)
    if (1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 25) != (
            1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 28):
        triggered.add(102)

    # 分支103-109: 调整条件和扩大范围
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1050 and 30 < moisture < 55 and (temp < 20 or temp > 22)):
        triggered.add(103)
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1150 and 22 < moisture < 55 and (temp < 20 or temp > 22)):
        triggered.add(104)
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1150 and 30 < moisture < 62 and (temp < 20 or temp > 22)):
        triggered.add(105)
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1150 and 30 < moisture < 55 and (temp < 16 or temp > 22)):
        triggered.add(106)
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 26)):
        triggered.add(107)
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1150 and 30 < moisture < 55 and (temp < 15 or temp > 22)):
        triggered.add(108)
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 28)):
        triggered.add(109)

    # 分支110-114: 调整到中间值
    if (moisture < 45 and co2 < 1150 and temp < 22) != (moisture < 38 and co2 < 1150 and temp < 22):
        triggered.add(110)
    if (moisture < 45 and co2 < 1150 and temp < 22) != (moisture < 35 and co2 < 1150 and temp < 22):
        triggered.add(111)
    if (moisture < 45 and co2 < 1150 and temp < 22) != (moisture < 45 and co2 < 1000 and temp < 22):
        triggered.add(112)
    if (moisture < 45 and co2 < 1150 and temp < 22) != (moisture < 45 and co2 < 1150 and temp < 27):
        triggered.add(113)
    if (moisture < 45 and co2 < 1150 and temp < 22) != (moisture < 45 and co2 < 1150 and temp < 30):
        triggered.add(114)

    # 分支115-118: 调整条件
    if (co2 < 1200 and moisture > 45 and temp > 20 and co2 < 1150) != (
            co2 < 1200 and moisture > 35 and temp > 20 and co2 < 1150):
        triggered.add(115)
    if (co2 < 1200 and moisture > 45 and temp > 20 and co2 < 1150) != (
            co2 < 1200 and moisture < 45 and temp > 20 and co2 < 1150):
        triggered.add(116)
    if (co2 < 1200 and moisture > 45 and temp > 20 and co2 < 1150) != (
            co2 < 1200 and moisture > 45 and temp > 15 and co2 < 1150):
        triggered.add(117)
    if (co2 < 1200 and moisture > 45 and temp > 20 and co2 < 1150) != (
            co2 < 1200 and moisture > 45 and temp > 20 and co2 < 1070):
        triggered.add(118)

    # 分支119-124: 调整到中间值和扩大范围
    if (co2 < 1150 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)) != (
            co2 < 1000 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)):
        triggered.add(119)
    if (co2 < 1150 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)) != (
            co2 < 1150 and (moisture > 30 or moisture < 40) and (temp > 22 or temp < 18)):
        triggered.add(120)
    if (co2 < 1150 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)) != (
            co2 < 1150 and (moisture > 40 or moisture < 40) and (temp > 22 or temp < 18)):
        triggered.add(121)
    if (co2 < 1150 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)) != (
            co2 < 1150 and (moisture > 50 and moisture < 40) and (temp > 22 or temp < 18)):
        triggered.add(122)
    if (co2 < 1150 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)) != (
            co2 < 1150 and (moisture > 50 or moisture < 45) and (temp > 22 or temp < 18)):
        triggered.add(123)
    if (co2 < 1150 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)) != (
            co2 < 1150 and (moisture > 50 or moisture < 40) or (temp > 22 or temp < 18)):
        triggered.add(124)

    # 分支125-131: 调整比例和阈值
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1050 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45):
        triggered.add(125)
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1100 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45):
        triggered.add(126)
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1150 and safe_divide(temp, moisture + 5) > 0.5 and moisture < 45):
        triggered.add(127)
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1150 and safe_divide(temp, moisture + 1) > 0.4 and moisture < 45):
        triggered.add(128)
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 or moisture < 45):
        triggered.add(129)
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 38):
        triggered.add(130)
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 35):
        triggered.add(131)

    # 分支132-137: 调整范围到中间
    if (co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 25) != (
            co2 < 1000 and 35 <= moisture <= 50 and 15 <= temp <= 25):
        triggered.add(132)
    if (co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 25) != (
            co2 < 1150 and 28 <= moisture <= 50 and 15 <= temp <= 25):
        triggered.add(133)
    if (co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 25) != (
            co2 < 1150 and 32 <= moisture <= 50 and 15 <= temp <= 25):
        triggered.add(134)
    if (co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 25) != (
            co2 < 1150 and 35 <= moisture <= 50 and 12 <= temp <= 25):
        triggered.add(135)
    if (co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 25) != (
            co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 28):
        triggered.add(136)
    if (co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 25) != (
            co2 < 1050 and 35 <= moisture <= 50 and 15 <= temp <= 25):
        triggered.add(137)

    return triggered


target_paths = [
    [1, 2, 6, 7, 8, 9, 14, 15, 16, 20, 24, 29, 30, 31, 35, 36, 40, 45, 47, 48, 54, 59, 60, 61, 68, 74, 75, 76, 86,
     87, 92, 93, 94, 103, 107, 109, 113, 114, 115, 116, 120, 121, 123, 124, 125, 127, 130, 131, 132, 137],
    [1, 2, 6, 7, 8, 9, 14, 16, 20, 24, 29, 30, 31, 35, 36, 40, 43, 44, 45, 47, 48, 54, 56, 57, 58, 68, 74, 75, 76, 84,
     85, 92, 93, 94, 103, 107, 109, 113, 114, 115, 116, 119, 122, 125, 130, 131, 132, 137],
    [1, 2, 6, 7, 8, 9, 14, 16, 20, 24, 29, 30, 31, 35, 36, 40, 43, 44, 45, 48, 54, 59, 64, 65, 68, 74, 76, 84, 85, 93,
     94, 102, 103, 107, 109, 113, 114, 115, 116, 119, 122, 125, 131, 136],
    [1, 2, 6, 7, 8, 9, 17, 18, 19, 20, 21, 29, 30, 31, 35, 36, 37, 38, 39, 42, 45, 47, 48, 54, 56, 57, 58, 68, 74, 75,
     76, 84, 85, 92, 93, 94, 111, 112, 115, 116, 124, 125, 131, 132, 137],
    [1, 2, 6, 7, 8, 9, 14, 15, 16, 21, 29, 30, 31, 35, 36, 40, 41, 42, 45, 47, 48, 54, 59, 60, 61, 68, 74, 75, 76, 86,
     87, 92, 93, 94, 110, 111, 112, 115, 116, 128, 129, 132, 137],
    [1, 2, 6, 14, 15, 16, 21, 36, 40, 41, 42, 45, 46, 47, 48, 59, 60, 61, 71, 72, 73, 74, 77, 78, 79, 86, 87, 92, 103,
     106, 108, 110, 111, 112, 120, 121, 123, 124, 129, 132, 137],
    [6, 7, 8, 9, 14, 15, 16, 20, 24, 29, 30, 31, 35, 36, 40, 48, 54, 59, 60, 61, 68, 74, 86, 87, 92, 93, 94, 98, 107,
     109, 113, 114, 115, 116, 120, 121, 123, 124, 127, 130, 131],
    [1, 2, 4, 9, 11, 14, 15, 16, 20, 24, 26, 27, 28, 29, 32, 40, 45, 47, 48, 54, 59, 68, 74, 75, 76, 86, 87, 91, 96,
     102, 103, 107, 109, 116, 118, 120, 121, 124, 129, 132, 137],
    [1, 2, 6, 7, 8, 9, 14, 15, 16, 24, 29, 30, 31, 35, 36, 40, 52, 55, 59, 68, 74, 75, 76, 86, 87, 93, 94, 103, 109,
     114, 115, 116, 120, 121, 123, 124, 125, 126, 130, 131, 136],
    [1, 2, 6, 8, 9, 14, 20, 24, 29, 30, 31, 35, 36, 40, 43, 44, 49, 50, 52, 59, 64, 65, 68, 74, 75, 76, 84, 85, 94,
     103, 107, 109, 113, 114, 116, 119, 122, 125, 131, 132, 137],
    [1, 2, 6, 8, 9, 17, 18, 19, 20, 21, 29, 30, 31, 35, 36, 37, 38, 39, 42, 49, 50, 52, 56, 57, 58, 68, 74, 75, 76,
     84, 85, 94, 99, 111, 112, 116, 124, 125, 131, 132, 137],
    [1, 2, 6, 10, 11, 17, 18, 19, 20, 21, 32, 33, 34, 35, 36, 40, 41, 42, 45, 47, 48, 59, 60, 61, 71, 72, 73, 74, 75,
     76, 86, 87, 88, 95, 96, 103, 106, 108, 117, 132, 137],
    [1, 2, 4, 5, 9, 11, 14, 15, 16, 24, 26, 27, 28, 29, 32, 40, 45, 46, 47, 48, 54, 59, 68, 74, 79, 86, 87, 91, 96,
     102, 103, 107, 109, 116, 118, 120, 121, 124, 129, 136],
    [2, 6, 7, 8, 9, 14, 16, 20, 24, 29, 30, 31, 35, 36, 40, 48, 54, 59, 60, 61, 68, 74, 86, 87, 92, 93, 94, 97, 98,
     107, 109, 113, 114, 115, 116, 120, 123, 124, 130, 131],
    [1, 2, 6, 10, 11, 14, 15, 16, 21, 32, 33, 34, 35, 36, 40, 41, 42, 45, 47, 48, 59, 60, 61, 71, 72, 73, 74, 79, 86,
     87, 88, 89, 95, 96, 103, 106, 108, 117, 132, 137],
    [1, 2, 6, 10, 11, 12, 13, 17, 32, 33, 34, 35, 36, 40, 42, 45, 46, 47, 48, 51, 59, 61, 71, 72, 73, 74, 77, 78, 79,
     87, 88, 89, 95, 96, 103, 106, 108, 117, 119, 122],
    [1, 2, 6, 8, 9, 14, 24, 29, 30, 31, 35, 36, 40, 43, 44, 49, 50, 52, 56, 57, 58, 68, 74, 79, 84, 85, 99, 103, 107,
     109, 113, 114, 116, 119, 122, 125, 126, 133, 134],
    [6, 7, 8, 9, 14, 15, 16, 20, 24, 29, 30, 31, 35, 36, 40, 48, 54, 59, 66, 67, 68, 74, 86, 87, 93, 94, 107, 109,
     113, 114, 115, 116, 120, 121, 123, 124, 130, 131],
    [1, 2, 6, 9, 14, 21, 22, 23, 24, 25, 29, 35, 36, 37, 38, 39, 42, 50, 52, 56, 57, 58, 66, 67, 69, 70, 71, 75, 76,
     84, 85, 112, 116, 124, 125, 126],
    [1, 2, 6, 10, 11, 14, 15, 16, 21, 32, 34, 35, 36, 40, 41, 42, 52, 53, 59, 71, 73, 74, 78, 79, 86, 87, 88, 89, 96,
     101, 103, 120, 121, 124, 135],
    [1, 2, 6, 14, 21, 36, 37, 38, 39, 42, 49, 50, 52, 56, 57, 58, 71, 72, 73, 74, 77, 78, 79, 84, 85, 99, 104, 112,
     119, 122, 125, 126, 127, 133],
    [1, 2, 4, 5, 9, 11, 12, 13, 17, 24, 26, 27, 28, 29, 32, 40, 45, 46, 47, 48, 51, 54, 59, 68, 74, 79, 88, 89, 96,
     100, 105, 116, 118, 119, 122],
    [1, 2, 6, 9, 14, 29, 35, 36, 40, 44, 59, 64, 65, 66, 67, 69, 70, 71, 75, 76, 80, 81, 84, 85, 104, 114, 116, 119,
     122, 125, 126],
    [1, 2, 6, 9, 14, 20, 24, 29, 30, 35, 36, 40, 44, 59, 64, 65, 68, 74, 75, 76, 80, 81, 82, 84, 85, 104, 114, 116,
     119, 122, 125],
    [1, 2, 6, 14, 16, 21, 36, 37, 38, 39, 42, 52, 53, 59, 62, 63, 71, 73, 74, 84, 85, 92, 101, 103, 111, 112, 119,
     122, 129, 135],
    [1, 2, 6, 10, 11, 14, 15, 16, 21, 32, 35, 36, 40, 41, 42, 52, 53, 59, 71, 74, 79, 86, 87, 90, 96, 103, 120, 121,
     124, 135],
    [1, 2, 6, 9, 14, 24, 29, 30, 35, 36, 40, 44, 59, 64, 65, 68, 74, 75, 76, 83, 84, 85, 104, 114, 116, 119, 122, 125,
     126],
    [3, 21, 22, 23, 24, 25, 35, 39, 40, 52, 58, 59, 66, 67, 69, 70, 71, 79, 94, 129]
]


def jaccard_similarity(set1: Set[int], set2: Set[int]) -> float:
    """If set1 covers target set2, the match score is treated as 1."""
    if set2.issubset(set1):
        return 1.0
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    return intersection / union if union != 0 else 0.0


def compute_reward(state, target_path, triggered):
    """Compute the DQN reward."""
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward


class SimpleReplayBuffer:
    def __init__(self, capacity=5000):
        self.buffer = deque(maxlen=capacity)

    def append(self, experience):
        self.buffer.append(experience)

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return random.sample(self.buffer, len(self.buffer))
        return random.sample(self.buffer, batch_size)

    def __len__(self):
        return len(self.buffer)


class SimpleDQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super().__init__()
        self.fc1 = nn.Linear(state_dim, 64)
        self.fc2 = nn.Linear(64, 32)
        self.fc3 = nn.Linear(32, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)


class SimpleDQNAgent:
    def __init__(self, state_dim, action_dim, learning_rate=0.001):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.epsilon = 0.3
        self.gamma = 0.99

        self.model = SimpleDQN(state_dim, action_dim).to(device)
        self.target_model = SimpleDQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())
        self.replay_buffer = SimpleReplayBuffer(capacity=5000)

    def decode_action(self, action_idx):
        """Decode the action."""
        light_deltas = [int((LIGHT_MAX - LIGHT_MIN) * p) for p in [0.05, 0.1, 0.2, 0.5, 0.7, -0.05, -0.1, -0.2, -0.5, -0.7]]
        moisture_deltas = [int((MOISTURE_MAX - MOISTURE_MIN) * p) for p in [0.05, 0.1, 0.2, 0.5, 0.7, -0.05, -0.1, -0.2, -0.5, -0.7]]
        temp_deltas = [int((TEMP_MAX - TEMP_MIN) * p) for p in [0.05, 0.1, 0.2, 0.5, 0.7, -0.05, -0.1, -0.2, -0.5, -0.7]]

        dim = action_idx // 10
        delta_idx = action_idx % 10

        if dim == 0:
            return (light_deltas[delta_idx], 0, 0)
        if dim == 1:
            return (0, moisture_deltas[delta_idx], 0)
        return (0, 0, temp_deltas[delta_idx])

    def store_transition(self, state, action, reward, next_state, done):
        self.replay_buffer.append((state, action, reward, next_state, done))

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return

        batch = self.replay_buffer.sample(batch_size)
        states, actions, rewards, next_states, dones = zip(*batch)

        states = torch.tensor(np.array(states), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array(next_states), dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))

        loss = nn.MSELoss()(current_q_values, target_q_values)

        self.optimizer.zero_grad()
        loss.backward()
        self.optimizer.step()

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())

    def get_best_samples(self, target_path, num_samples=20):
        """Screen DQN candidate samples from the replay buffer. This stage is counted separately in T_sample_screening."""
        if len(self.replay_buffer) == 0:
            return []

        samples_with_scores = []
        for state_norm, _, _, _, _ in self.replay_buffer.buffer:
            state_original = denormalize_state(state_norm)
            triggered = execute_Tr(state_original)
            reward = compute_reward(state_original, target_path, triggered)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_scores.append((state_original, reward, sim, triggered))

        samples_with_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_scores[:num_samples]


def train_dqn_for_path(path_idx, target_path, num_samples=200):
    """Train a standard DQN for a single path."""
    print(f"  Start training path {path_idx + 1} DQN model...")
    agent = SimpleDQNAgent(state_dim=3, action_dim=30)

    random_states = [normalize_state(generate_random_state()) for _ in range(num_samples)]

    STEPS_PER_SAMPLE = 3
    EPOCHS = 5
    BATCH_SIZE = 32

    step_count = 0
    for _ in range(EPOCHS):
        for state_norm in random_states:
            state_norm = tuple(state_norm)
            state_original = denormalize_state(state_norm)

            for step in range(STEPS_PER_SAMPLE):
                legal_actions = []
                for a in range(agent.action_dim):
                    dx, dy, dz = agent.decode_action(a)
                    next_state_candidate = (state_original[0] + dx, state_original[1] + dy, state_original[2] + dz)
                    if is_state_valid(next_state_candidate):
                        legal_actions.append(a)

                if not legal_actions:
                    break

                if random.random() < agent.epsilon:
                    action = random.choice(legal_actions)
                else:
                    state_tensor = torch.tensor(state_norm, dtype=torch.float32).unsqueeze(0).to(device)
                    with torch.no_grad():
                        q_values = agent.model(state_tensor)[0]
                    action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                dx, dy, dz = agent.decode_action(action)
                next_state_original = (state_original[0] + dx, state_original[1] + dy, state_original[2] + dz)
                next_state_norm = normalize_state(next_state_original)

                triggered = execute_Tr(next_state_original)
                reward = compute_reward(next_state_original, target_path, triggered)
                done = (step == STEPS_PER_SAMPLE - 1)

                agent.store_transition(state_norm, action, reward, next_state_norm, done)

                state_norm = next_state_norm
                state_original = next_state_original
                step_count += 1

                if step_count % 50 == 0 and len(agent.replay_buffer) >= BATCH_SIZE:
                    agent.train(BATCH_SIZE)

                if step_count % 100 == 0:
                    agent.update_target_model()

    print(f"  Path {path_idx + 1} training completed, replay buffer size: {len(agent.replay_buffer)}")
    return agent


class Particle:
    def __init__(self, initial_position=None):
        if initial_position is not None:
            self.position = clip_state(np.array(initial_position, dtype=float))
        else:
            self.position = np.array([
                np.random.uniform(BOUNDS["light"][0], BOUNDS["light"][1]),
                np.random.uniform(BOUNDS["moisture"][0], BOUNDS["moisture"][1]),
                np.random.uniform(BOUNDS["temp"][0], BOUNDS["temp"][1]),
            ], dtype=float)

        spans = np.array([
            BOUNDS["light"][1] - BOUNDS["light"][0],
            BOUNDS["moisture"][1] - BOUNDS["moisture"][0],
            BOUNDS["temp"][1] - BOUNDS["temp"][0],
        ], dtype=float)
        self.velocity = np.random.uniform(-0.1 * spans, 0.1 * spans)
        self.best_position = self.position.copy()
        self.best_fitness = -1.0
        self.fitness = -1.0


class BasicPSO:
    """
    Standard PSO: contains only standard velocity and position updates.
    Mutation coefficients, opposite particles, local-best reset, and other improved PSO mechanisms have been removed.
    """

    def __init__(self, target_path, swarm_size=20, dqn_samples=None):
        self.target_path = target_path
        self.swarm_size = swarm_size
        self.particles = []
        self.global_best_position = None
        self.global_best_fitness = -1.0

        if dqn_samples is not None and len(dqn_samples) > 0:
            num_direct = min(len(dqn_samples), swarm_size)
            for i in range(num_direct):
                state_tuple, _, _, _ = dqn_samples[i]
                self.particles.append(Particle(initial_position=state_tuple))

            while len(self.particles) < swarm_size:
                base_idx = len(self.particles) % len(dqn_samples)
                state_tuple, _, _, _ = dqn_samples[base_idx]
                perturbed = np.array(state_tuple, dtype=float) + np.random.randint(-10, 11, size=3)
                self.particles.append(Particle(initial_position=clip_state(perturbed)))
        else:
            self.particles = [Particle() for _ in range(swarm_size)]

        for particle in self.particles:
            particle.fitness = self.fitness_function(particle.position)
            particle.best_fitness = particle.fitness
            particle.best_position = particle.position.copy()

            if particle.fitness > self.global_best_fitness:
                self.global_best_fitness = particle.fitness
                self.global_best_position = particle.position.copy()

    def fitness_function(self, position):
        try:
            triggered = execute_Tr(position)
            return jaccard_similarity(triggered, self.target_path)
        except Exception:
            return 0.0

    def update(self):
        spans = np.array([
            BOUNDS["light"][1] - BOUNDS["light"][0],
            BOUNDS["moisture"][1] - BOUNDS["moisture"][0],
            BOUNDS["temp"][1] - BOUNDS["temp"][0],
        ], dtype=float)
        max_velocity = PSO_VMAX_RATIO * spans

        for particle in self.particles:
            r1 = np.random.random(3)
            r2 = np.random.random(3)

            particle.velocity = (
                PSO_W * particle.velocity
                + PSO_C1 * r1 * (particle.best_position - particle.position)
                + PSO_C2 * r2 * (self.global_best_position - particle.position)
            )
            particle.velocity = np.clip(particle.velocity, -max_velocity, max_velocity)

            particle.position = clip_state(particle.position + particle.velocity)
            particle.fitness = self.fitness_function(particle.position)

            if particle.fitness > particle.best_fitness:
                particle.best_fitness = particle.fitness
                particle.best_position = particle.position.copy()

            if particle.fitness > self.global_best_fitness:
                self.global_best_fitness = particle.fitness
                self.global_best_position = particle.position.copy()


def _set_header(ws, headers, widths, header_fill, header_font, border, center_align):
    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_row(ws, row_idx, row_data, border, alignment, fill=None):
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.border = border
        cell.alignment = alignment
        if fill is not None:
            cell.fill = fill


def export_time_metrics_to_excel(all_run_results, experiment_total_time=None, filename=None):
    """Export only timing metrics and per-path iteration counts."""
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"DQN_BasicPSO_Time_Metrics_{timestamp}.xlsx"

    output_dir = os.path.join(os.getcwd(), "results")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    num_runs = len(all_run_results)
    num_paths = len(target_paths)

    # Worksheet 1: runtime summary
    ws1 = wb.active
    ws1.title = "Runtime Summary"
    ws1.sheet_view.showGridLines = False
    headers1 = [
        "Run", "Number of Paths", "DQNTraining Total Time(s)", "Sample Screening Total Time(s)", "DQNDirect Check Total Time(s)",
        "PSOInitialization Total Time(s)", "PSOSearch Total Time(s)", "Total Algorithm Time(s)", "Total Iterations", "Average Iterations per Path"
    ]
    widths1 = [12, 10, 18, 18, 22, 18, 18, 16, 14, 18]
    _set_header(ws1, headers1, widths1, header_fill, header_font, border, center_align)

    for run_idx, run_result in enumerate(all_run_results, 1):
        path_results = run_result["paths"]
        total_dqn_train = sum(p["T_DQN_train"] for p in path_results)
        total_screen = sum(p["T_sample_screening"] for p in path_results)
        total_direct_check = sum(p["T_direct_check"] for p in path_results)
        total_pso_init = sum(p["T_PSO_init"] for p in path_results)
        total_pso_search = sum(p["T_PSO_search"] for p in path_results)
        total_iterations = sum(p["iterations"] for p in path_results)

        row_data = [
            f" {run_idx}",
            num_paths,
            round(total_dqn_train, 6),
            round(total_screen, 6),
            round(total_direct_check, 6),
            round(total_pso_init, 6),
            round(total_pso_search, 6),
            round(run_result["T_run_algorithm_total"], 6),
            int(total_iterations),
            round(total_iterations / num_paths, 2),
        ]
        _write_row(ws1, run_idx + 1, row_data, border, center_align, alternate_fill if run_idx % 2 == 0 else None)

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:J{num_runs + 1}"

    # Worksheet 2: path timing and iteration details
    ws2 = wb.create_sheet(title="Path Timing and Iteration Details")
    ws2.sheet_view.showGridLines = False
    headers2 = [
        "Run", "Path ID", "DQN training(s)", "Sample Screening Time(s)", "DQNDirect Check Time(s)",
        "PSOInitialization Time(s)", "PSOSearch Time(s)", "Path Total Algorithm Time(s)", "Iterations"
    ]
    widths2 = [12, 12, 18, 18, 22, 18, 18, 18, 12]
    _set_header(ws2, headers2, widths2, header_fill, header_font, border, center_align)

    row_idx = 2
    for run_idx, run_result in enumerate(all_run_results, 1):
        for path_result in run_result["paths"]:
            row_data = [
                f" {run_idx}",
                f"Path  {path_result['path_idx'] + 1}",
                round(path_result["T_DQN_train"], 6),
                round(path_result["T_sample_screening"], 6),
                round(path_result["T_direct_check"], 6),
                round(path_result["T_PSO_init"], 6),
                round(path_result["T_PSO_search"], 6),
                round(path_result["T_path_algorithm_total"], 6),
                int(path_result["iterations"]),
            ]
            _write_row(ws2, row_idx, row_data, border, center_align, alternate_fill if row_idx % 2 == 0 else None)
            row_idx += 1

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:I{row_idx - 1}"

    # Worksheet 3: path aggregate statistics
    ws3 = wb.create_sheet(title="Path Aggregate Statistics")
    ws3.sheet_view.showGridLines = False
    headers3 = [
        "Path ID", "Average DQN Training Time(s)", "Average Sample Screening Time(s)", "Average PSO Initialization Time(s)",
        "Average PSO Search Time(s)", "Average Path Total Algorithm Time(s)", "Average Iterations", "Minimum Iterations", "Maximum Iterations"
    ]
    widths3 = [12, 22, 22, 22, 22, 22, 16, 14, 14]
    _set_header(ws3, headers3, widths3, header_fill, header_font, border, center_align)

    for path_idx in range(num_paths):
        records = [run_result["paths"][path_idx] for run_result in all_run_results]
        iterations = [r["iterations"] for r in records]
        row_data = [
            f"Path  {path_idx + 1}",
            round(float(np.mean([r["T_DQN_train"] for r in records])), 6),
            round(float(np.mean([r["T_sample_screening"] for r in records])), 6),
            round(float(np.mean([r["T_PSO_init"] for r in records])), 6),
            round(float(np.mean([r["T_PSO_search"] for r in records])), 6),
            round(float(np.mean([r["T_path_algorithm_total"] for r in records])), 6),
            round(float(np.mean(iterations)), 2),
            int(np.min(iterations)),
            int(np.max(iterations)),
        ]
        _write_row(ws3, path_idx + 2, row_data, border, center_align, alternate_fill if (path_idx + 1) % 2 == 0 else None)

    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:I{num_paths + 1}"

    # Worksheet 4: experiment overview
    ws4 = wb.create_sheet(title="Experiment Overview")
    ws4.sheet_view.showGridLines = False
    headers4 = ["Metric", "Value"]
    widths4 = [32, 24]
    _set_header(ws4, headers4, widths4, header_fill, header_font, border, center_align)

    all_path_records = [p for run_result in all_run_results for p in run_result["paths"]]
    all_iterations = [p["iterations"] for p in all_path_records]
    overview_rows = [
        ["Experiment Runs", num_runs],
        ["Number of Target Paths", num_paths],
        ["Total Experiment Wall-clock Time(s)", round(experiment_total_time, 6) if experiment_total_time is not None else ""],
        ["Average Total Algorithm Time per Run(s)", round(float(np.mean([r["T_run_algorithm_total"] for r in all_run_results])), 6)],
        ["Average Total Algorithm Time per Path(s)", round(float(np.mean([p["T_path_algorithm_total"] for p in all_path_records])), 6)],
        ["Average DQN Training Time per Path(s)", round(float(np.mean([p["T_DQN_train"] for p in all_path_records])), 6)],
        ["Average Sample Screening Time per Path(s)", round(float(np.mean([p["T_sample_screening"] for p in all_path_records])), 6)],
        ["Average PSO Search Time per Path(s)", round(float(np.mean([p["T_PSO_search"] for p in all_path_records])), 6)],
        ["Average Iterations per Path", round(float(np.mean(all_iterations)), 2)],
        ["Maximum Iterations per Path", int(np.max(all_iterations))],
        ["Minimum Iterations per Path", int(np.min(all_iterations))],
    ]

    for idx, row_data in enumerate(overview_rows, 2):
        _write_row(ws4, idx, row_data, border, center_align, alternate_fill if idx % 2 == 0 else None)

    wb.save(filepath)
    print(f"\n Timing metrics exported to: {filepath}")
    return filepath


def run_single_experiment(run_num, max_iterations=3000):
    """Run one DQN + standard PSO ."""
    print(f"\n{'=' * 80}")
    print(f"Start run  {run_num}  run: DQN + standard PSO ")
    print(f"{'=' * 80}")

    run_start = time.perf_counter()
    path_results = []

    for path_idx, target_path in enumerate(target_paths):
        print(f"\nRun {run_num} run - Path {path_idx + 1}")
        path_total_start = time.perf_counter()

        # 1) DQN training
        dqn_train_start = time.perf_counter()
        agent = train_dqn_for_path(path_idx, target_path, num_samples=200)
        T_DQN_train = time.perf_counter() - dqn_train_start

        # 2) DQNSample Screening Time
        screening_start = time.perf_counter()
        dqn_samples = agent.get_best_samples(target_path, num_samples=20)
        T_sample_screening = time.perf_counter() - screening_start

        # 3) DQN
        direct_check_start = time.perf_counter()
        direct_solution_found = False
        for state_tuple, reward, sim, triggered in dqn_samples:
            if sim >= 1.0 or target_path.issubset(triggered):
                direct_solution_found = True
                break
        T_direct_check = time.perf_counter() - direct_check_start

        T_PSO_init = 0.0
        T_PSO_search = 0.0
        iterations_used = 0

        if direct_solution_found:
            print(
                f"  Path {path_idx + 1}: DQN | "
                f"DQN training {T_DQN_train:.6f}s | screening {T_sample_screening:.6f}s | iterations {iterations_used}"
            )
        else:
            # 4) standard PSOInitialization Time
            pso_init_start = time.perf_counter()
            pso = BasicPSO(target_path, swarm_size=20, dqn_samples=dqn_samples)
            T_PSO_init = time.perf_counter() - pso_init_start

            # 5) standard PSOSearch Time
            pso_search_start = time.perf_counter()
            iterations_used = max_iterations
            for iteration in range(max_iterations):
                pso.update()
                if pso.global_best_fitness >= 1.0:
                    iterations_used = iteration + 1
                    break
            T_PSO_search = time.perf_counter() - pso_search_start

            print(
                f"  Path {path_idx + 1}: "
                f"DQN training {T_DQN_train:.6f}s | screening {T_sample_screening:.6f}s | "
                f"PSO {T_PSO_init:.6f}s | PSO {T_PSO_search:.6f}s | "
                f"iterations {iterations_used}"
            )

        T_path_algorithm_total = time.perf_counter() - path_total_start

        path_results.append({
            "path_idx": path_idx,
            "T_DQN_train": T_DQN_train,
            "T_sample_screening": T_sample_screening,
            "T_direct_check": T_direct_check,
            "T_PSO_init": T_PSO_init,
            "T_PSO_search": T_PSO_search,
            "T_path_algorithm_total": T_path_algorithm_total,
            "iterations": iterations_used,
        })

    T_run_algorithm_total = time.perf_counter() - run_start
    print(f"\nRun {run_num} runcompleted | Total Algorithm Time {T_run_algorithm_total:.6f}s")
    return {
        "run_num": run_num,
        "paths": path_results,
        "T_run_algorithm_total": T_run_algorithm_total,
    }


def run_multiple_experiments(num_runs):
    print("\n" + "=" * 80)
    print(f"DQN + standard PSO - {num_runs}: Metric")
    print("=" * 80)
    print(f"Number of Target Paths: {len(target_paths)}")
    print("Statistics: DQN training, Sample Screening Time, PSOInitialization Time, PSOSearch Time, Path Iterations")
    print("=" * 80)

    all_run_results = []
    experiment_start = time.perf_counter()

    for run_num in range(1, num_runs + 1):
        run_result = run_single_experiment(run_num)
        all_run_results.append(run_result)

    experiment_total_time = time.perf_counter() - experiment_start
    print(f"\n{'=' * 80}")
    print(f"All {num_runs} runcompleted | Total Experiment Wall-clock Time {experiment_total_time:.6f}s")
    print(f"{'=' * 80}\n")
    return all_run_results, experiment_total_time


if __name__ == "__main__":
    print("=" * 80)
    print("DQN + standard PSO - MetricPath Iterations")
    print("=" * 80)
    print(f"Current configuration: Run = {NUM_RUNS}")
    print(f"Number of Paths: {len(target_paths)}")
    print(f"Device: {device}")
    print("=" * 80)

    if len(sys.argv) > 1:
        try:
            NUM_RUNS = int(sys.argv[1])
            print(f"Read from command line: Run = {NUM_RUNS}")
        except ValueError:
            print(f"Invalid command-line argument, using default number of runs {NUM_RUNS}")

    all_results, total_time = run_multiple_experiments(num_runs=NUM_RUNS)
    export_time_metrics_to_excel(all_results, total_time)
    print("\nProgram completed")
