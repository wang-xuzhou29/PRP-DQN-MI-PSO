import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

def execute_validation_rules(moisture, co2, temp):
    """检测低CO2极端条件下的分支触发情况"""
    triggered = set()

    # Fixed: proper indentation and using triggered.add()
    if (co2 < 1150) != (co2 < 1000):
        triggered.add(1)
    if (co2 < 1150) != (co2 < 950):
        triggered.add(2)
    if (co2 < 1150) != (co2 < 1400):
        triggered.add(3)

    # 分支4-11: 调整到中间值区域
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1050 and moisture > 45 and temp > 20):
        triggered.add(4)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1100 and moisture > 45 and temp > 20):
        triggered.add(5)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 or moisture > 45 and temp > 20):
        triggered.add(6)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 35 and temp > 20):
        triggered.add(7)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 30 and temp > 20):
        triggered.add(8)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture < 45 and temp > 20):
        triggered.add(9)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 45 and temp > 5):
        triggered.add(10)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 45 and temp < 20):
        triggered.add(11)

    # 分支12-21: 调整到中间值
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 970 and moisture > 50) or (co2 < 1100 and temp < 22)):
        triggered.add(12)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1000 and moisture > 50) or (co2 < 1100 and temp < 22)):
        triggered.add(13)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 or moisture > 50) or (co2 < 1100 and temp < 22)):
        triggered.add(14)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 40) or (co2 < 1100 and temp < 22)):
        triggered.add(15)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 35) or (co2 < 1100 and temp < 22)):
        triggered.add(16)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 50) and (co2 < 1100 and temp < 22)):
        triggered.add(17)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 50) or (co2 < 960 and temp < 22)):
        triggered.add(18)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 18)):
        triggered.add(19)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 50) or (co2 < 1100 and temp > 22)):
        triggered.add(20)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 50) or (co2 > 1100 and temp < 22)):
        triggered.add(21)

    # 分支22-25: 调整阈值到中间 - using safe_divide
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp < 22) != (safe_divide(moisture, co2 - 700) > 0.04 and temp < 22):
        triggered.add(22)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp < 22) != (safe_divide(moisture, co2 - 700) > 0.03 and temp < 22):
        triggered.add(23)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp < 22) != (safe_divide(moisture, co2 - 700) > 0.06 or temp < 22):
        triggered.add(24)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp < 22) != (moisture + safe_divide(100, co2 - 700) > 0.06 and temp < 22):
        triggered.add(25)

    # 分支26-36: 调整到中间值
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1000 and moisture > 45 and temp > 20):
        triggered.add(26)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1070 and moisture > 45 and temp > 20):
        triggered.add(27)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 > 1150 and moisture > 45 and temp > 20):
        triggered.add(28)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture < 45 and temp > 20):
        triggered.add(29)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 25 and temp > 20):
        triggered.add(30)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 30 and temp > 20):
        triggered.add(31)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 45 and temp < 20):
        triggered.add(32)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 45 and temp > 15):
        triggered.add(33)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 45 and temp > 12):
        triggered.add(34)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 45 or temp > 20):
        triggered.add(35)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 or moisture > 45 and temp > 20):
        triggered.add(36)

    # 分支37-44: 调整到中间值
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 970 and moisture < 40 and temp < 22):
        triggered.add(37)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 1020 and moisture < 40 and temp < 22):
        triggered.add(38)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 > 1150 and moisture < 40 and temp < 22):
        triggered.add(39)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 1150 or moisture < 40 and temp < 22):
        triggered.add(40)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 1150 and moisture < 50 and temp < 22):
        triggered.add(41)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 1150 and moisture > 40 and temp < 22):
        triggered.add(42)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 1150 and moisture < 40 and temp < 27):
        triggered.add(43)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 1150 and moisture < 40 and temp < 32):
        triggered.add(44)

    # 分支45-55: 扩大范围
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1050 and 35 < moisture < 60 and 15 < temp < 28):
        triggered.add(45)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1120 and 35 < moisture < 60 and 15 < temp < 28):
        triggered.add(46)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1070 and 35 < moisture < 60 and 15 < temp < 28):
        triggered.add(47)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 > 1150 and 35 < moisture < 60 and 15 < temp < 28):
        triggered.add(48)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 25 < moisture < 60 and 15 < temp < 28):
        triggered.add(49)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 20 < moisture < 60 and 15 < temp < 28):
        triggered.add(50)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 35 < moisture < 52 and 15 < temp < 28):
        triggered.add(51)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 35 < moisture < 60 or 15 < temp < 28):
        triggered.add(52)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 35 < moisture < 60 and 10 < temp < 28):
        triggered.add(53)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 35 < moisture < 60 and 15 < temp < 20):
        triggered.add(54)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 35 < moisture < 60 and 15 < temp < 33):
        triggered.add(55)

    # 分支56-65: 调整到中间值和扩大范围
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1020 and moisture < 40 and 15 < temp < 25):
        triggered.add(56)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1070 and moisture < 40 and 15 < temp < 25):
        triggered.add(57)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 > 1150 and moisture < 40 and 15 < temp < 25):
        triggered.add(58)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 or moisture < 40 and 15 < temp < 25):
        triggered.add(59)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 and moisture < 50 and 15 < temp < 25):
        triggered.add(60)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 and moisture < 55 and 15 < temp < 25):
        triggered.add(61)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 and moisture < 40 and 10 < temp < 25):
        triggered.add(62)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 and moisture < 40 and 5 < temp < 25):
        triggered.add(63)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 and moisture < 40 and 15 < temp < 30):
        triggered.add(64)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 and moisture < 40 and 15 < temp < 30):
        triggered.add(65)

    # 分支66-74: 调整比例和温度阈值
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 850) > 0.06 and temp > 20):
        triggered.add(66)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 900) > 0.06 and temp > 20):
        triggered.add(67)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 + 200) > 0.06 and temp > 20):
        triggered.add(68)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 700) > 0.04 and temp > 20):
        triggered.add(69)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 700) > 0.03 and temp > 20):
        triggered.add(70)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 700) > 0.06 or temp > 20):
        triggered.add(71)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 700) > 0.06 and temp > 15):
        triggered.add(72)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 700) > 0.06 and temp > 12):
        triggered.add(73)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 700) > 0.06 and temp < 20):
        triggered.add(74)

    if (co2 + moisture > 1150 and temp > 18) != (co2 + moisture > 1100 and temp > 18):
        triggered.add(75)
    if (co2 + moisture > 1150 and temp > 18) != (co2 + moisture > 1050 and temp > 18):
        triggered.add(76)
    if (co2 + moisture > 1150 and temp > 18) != (co2 + moisture > 1150 and temp > 15):
        triggered.add(77)
    if (co2 + moisture > 1150 and temp > 18) != (co2 + moisture > 1150 and temp > 12):
        triggered.add(78)
    if (co2 + moisture > 1150 and temp > 18) != (co2 + moisture > 1150 and temp < 18):
        triggered.add(79)

    if (co2 < 1150 and abs(moisture - 45) > 15 and abs(temp - 20) > 7) != (
            co2 < 1000 and abs(moisture - 45) > 15 and abs(temp - 20) > 7):
        triggered.add(80)
    if (co2 < 1150 and abs(moisture - 45) > 15 and abs(temp - 20) > 7) != (
            co2 < 1050 and abs(moisture - 45) > 15 and abs(temp - 20) > 7):
        triggered.add(81)
    if (co2 < 1150 and abs(moisture - 45) > 15 and abs(temp - 20) > 7) != (
            co2 < 1150 and abs(moisture - 42) > 15 and abs(temp - 20) > 7):
        triggered.add(82)
    if (co2 < 1150 and abs(moisture - 45) > 15 and abs(temp - 20) > 7) != (
            co2 < 1150 and abs(moisture - 45) > 12 and abs(temp - 20) > 7):
        triggered.add(83)

    # 分支84-87: 简化条件到中间值
    if (co2 < 1150 and moisture < 40) != (co2 < 1000 and moisture < 40):
        triggered.add(84)
    if (co2 < 1150 and moisture < 40) != (co2 < 1050 and moisture < 40):
        triggered.add(85)
    if (co2 < 1150 and moisture < 40) != (co2 < 1150 and moisture < 50):
        triggered.add(86)
    if (co2 < 1150 and moisture < 40) != (co2 < 1150 and moisture < 55):
        triggered.add(87)

    # 分支88-92: 调整到中间范围
    if (co2 < 1150 and 12 < temp < 25 and moisture > 45) != (co2 < 1030 and 12 < temp < 25 and moisture > 45):
        triggered.add(88)
    if (co2 < 1150 and 12 < temp < 25 and moisture > 45) != (co2 < 1080 and 12 < temp < 25 and moisture > 45):
        triggered.add(89)
    if (co2 < 1150 and 12 < temp < 25 and moisture > 45) != (co2 < 1150 and 8 < temp < 25 and moisture > 45):
        triggered.add(90)
    if (co2 < 1150 and 12 < temp < 25 and moisture > 45) != (co2 < 1150 and 12 < temp < 30 and moisture > 45):
        triggered.add(91)
    if (co2 < 1150 and 12 < temp < 25 and moisture > 45) != (co2 < 1150 and 12 < temp < 25 and moisture > 35):
        triggered.add(92)

    # 分支93-96: 调整到中间值
    if (moisture > 45 and temp > 20) != (moisture > 35 and temp > 20):
        triggered.add(93)
    if (moisture > 45 and temp > 20) != (moisture > 32 and temp > 20):
        triggered.add(94)
    if (moisture > 45 and temp > 20) != (moisture > 45 and temp > 15):
        triggered.add(95)
    if (moisture > 45 and temp > 20) != (moisture > 45 and temp < 20):
        triggered.add(96)

    # 分支97-102: 扩大范围
    if (1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 25) != (
            950 < co2 < 1250 and 35 < moisture < 55 and 15 < temp < 25):
        triggered.add(97)
    if (1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 25) != (
            900 < co2 < 1300 and 35 < moisture < 55 and 15 < temp < 25):
        triggered.add(98)
    if (1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 25) != (
            1000 < co2 < 1200 and 28 < moisture < 55 and 15 < temp < 25):
        triggered.add(99)
    if (1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 25) != (
            1000 < co2 < 1200 and 35 < moisture < 62 and 15 < temp < 25):
        triggered.add(100)
    if (1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 25) != (
            1000 < co2 < 1200 and 35 < moisture < 55 and 12 < temp < 25):
        triggered.add(101)
    if (1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 25) != (
            1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 28):
        triggered.add(102)

    # 分支103-109: 调整条件和扩大范围
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1050 and 30 < moisture < 55 and (temp < 20 or temp > 22)):
        triggered.add(103)
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1150 and 22 < moisture < 55 and (temp < 20 or temp > 22)):
        triggered.add(104)
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1150 and 30 < moisture < 62 and (temp < 20 or temp > 22)):
        triggered.add(105)
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1150 and 30 < moisture < 55 and (temp < 16 or temp > 22)):
        triggered.add(106)
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 26)):
        triggered.add(107)
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1150 and 30 < moisture < 55 and (temp < 15 or temp > 22)):
        triggered.add(108)
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 28)):
        triggered.add(109)

    # 分支110-114: 调整到中间值
    if (moisture < 45 and co2 < 1150 and temp < 22) != (moisture < 38 and co2 < 1150 and temp < 22):
        triggered.add(110)
    if (moisture < 45 and co2 < 1150 and temp < 22) != (moisture < 35 and co2 < 1150 and temp < 22):
        triggered.add(111)
    if (moisture < 45 and co2 < 1150 and temp < 22) != (moisture < 45 and co2 < 1000 and temp < 22):
        triggered.add(112)
    if (moisture < 45 and co2 < 1150 and temp < 22) != (moisture < 45 and co2 < 1150 and temp < 27):
        triggered.add(113)
    if (moisture < 45 and co2 < 1150 and temp < 22) != (moisture < 45 and co2 < 1150 and temp < 30):
        triggered.add(114)

    # 分支115-118: 调整条件
    if (co2 < 1200 and moisture > 45 and temp > 20 and co2 < 1150) != (
            co2 < 1200 and moisture > 35 and temp > 20 and co2 < 1150):
        triggered.add(115)
    if (co2 < 1200 and moisture > 45 and temp > 20 and co2 < 1150) != (
            co2 < 1200 and moisture < 45 and temp > 20 and co2 < 1150):
        triggered.add(116)
    if (co2 < 1200 and moisture > 45 and temp > 20 and co2 < 1150) != (
            co2 < 1200 and moisture > 45 and temp > 15 and co2 < 1150):
        triggered.add(117)
    if (co2 < 1200 and moisture > 45 and temp > 20 and co2 < 1150) != (
            co2 < 1200 and moisture > 45 and temp > 20 and co2 < 1070):
        triggered.add(118)

    # 分支119-124: 调整到中间值和扩大范围
    if (co2 < 1150 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)) != (
            co2 < 1000 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)):
        triggered.add(119)
    if (co2 < 1150 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)) != (
            co2 < 1150 and (moisture > 30 or moisture < 40) and (temp > 22 or temp < 18)):
        triggered.add(120)
    if (co2 < 1150 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)) != (
            co2 < 1150 and (moisture > 40 or moisture < 40) and (temp > 22 or temp < 18)):
        triggered.add(121)
    if (co2 < 1150 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)) != (
            co2 < 1150 and (moisture > 50 and moisture < 40) and (temp > 22 or temp < 18)):
        triggered.add(122)
    if (co2 < 1150 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)) != (
            co2 < 1150 and (moisture > 50 or moisture < 45) and (temp > 22 or temp < 18)):
        triggered.add(123)
    if (co2 < 1150 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)) != (
            co2 < 1150 and (moisture > 50 or moisture < 40) or (temp > 22 or temp < 18)):
        triggered.add(124)

    # 分支125-131: 调整比例和阈值
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1050 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45):
        triggered.add(125)
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1100 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45):
        triggered.add(126)
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1150 and safe_divide(temp, moisture + 5) > 0.5 and moisture < 45):
        triggered.add(127)
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1150 and safe_divide(temp, moisture + 1) > 0.4 and moisture < 45):
        triggered.add(128)
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 or moisture < 45):
        triggered.add(129)
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 38):
        triggered.add(130)
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 35):
        triggered.add(131)

    # 分支132-137: 调整范围到中间
    if (co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 25) != (
            co2 < 1000 and 35 <= moisture <= 50 and 15 <= temp <= 25):
        triggered.add(132)
    if (co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 25) != (
            co2 < 1150 and 28 <= moisture <= 50 and 15 <= temp <= 25):
        triggered.add(133)
    if (co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 25) != (
            co2 < 1150 and 32 <= moisture <= 50 and 15 <= temp <= 25):
        triggered.add(134)
    if (co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 25) != (
            co2 < 1150 and 35 <= moisture <= 50 and 12 <= temp <= 25):
        triggered.add(135)
    if (co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 25) != (
            co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 28):
        triggered.add(136)
    if (co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 25) != (
            co2 < 1050 and 35 <= moisture <= 50 and 15 <= temp <= 25):
        triggered.add(137)

    return triggered


def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """"""
    generated_path = execute_validation_rules(particle[0], particle[1], particle[2])

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # : x:1-50, y:1-50, z:1-50
        self.bounds = bounds if bounds else [(1, 100), (1, 100), (1, 100)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """target pathsPSO"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(particles[i][0], particles[i][1], particles[i][2]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(gbest_particle[0], gbest_particle[1], gbest_particle[2]),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """Path PSO"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - Path ")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations")
    print(f"Path : {len(target_paths)}")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"Path {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "" if result['success'] else f"({result['best_fitness']:.3f})"
        print(f"{status} | {result['time']:.2f}s | iterations{result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f": {success_count}/{len(target_paths)} ({success_rate:.1f}%) | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs}")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations, {len(target_paths)}Path ")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"---  {run_idx}/{num_runs} ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f": {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs} runcompleted | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 1: 
    ws1 = wb.active
    ws1.title = ""
    ws1.sheet_view.showGridLines = False

    headers = ["Run", "", "", "", "Average Iterations", "(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f" {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # 2: Path 
    ws2 = wb.create_sheet(title="Path ")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Path ID", "", "", "", "Average Iterations", "Minimum Iterations", "Maximum Iterations"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"Path  {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # 3: 
    ws3 = wb.create_sheet(title="")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Path ", "", "(x,y,z)", "", "Iterations", "Path "]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"{run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # 4: target paths
    ws4 = wb.create_sheet(title="target paths")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Path ID", "target paths", ""]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"Path  {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f" : {filename}")
    print(f"{'=' * 70}")
    print(f":")
    print(f"  1.        - {len(all_results)} run")
    print(f"  2. Path        - Path ")
    print(f"  3.    -  runPath ")
    print(f"  4. target paths       - target paths")
    print(f"{'=' * 70}\n")

    return filename


def main():


    target_paths = [
        [1, 2, 6, 7, 8, 9, 14, 15, 16, 20, 24, 29, 30, 31, 35, 36, 40, 45, 47, 48, 54, 59, 60, 61, 68, 74, 75, 76, 86,
         87, 92, 93, 94, 103, 107, 109, 113, 114, 115, 116, 120, 121, 123, 124, 125, 127, 130, 131, 132, 137],
        [1, 2, 6, 7, 8, 9, 14, 16, 20, 24, 29, 30, 31, 35, 36, 40, 43, 44, 45, 47, 48, 54, 56, 57, 58, 68, 74, 75, 76,
         84,
         85, 92, 93, 94, 103, 107, 109, 113, 114, 115, 116, 119, 122, 125, 130, 131, 132, 137],
        [1, 2, 6, 7, 8, 9, 14, 16, 20, 24, 29, 30, 31, 35, 36, 40, 43, 44, 45, 48, 54, 59, 64, 65, 68, 74, 76, 84, 85,
         93,
         94, 102, 103, 107, 109, 113, 114, 115, 116, 119, 122, 125, 131, 136],
        [1, 2, 6, 7, 8, 9, 17, 18, 19, 20, 21, 29, 30, 31, 35, 36, 37, 38, 39, 42, 45, 47, 48, 54, 56, 57, 58, 68, 74,
         75,
         76, 84, 85, 92, 93, 94, 111, 112, 115, 116, 124, 125, 131, 132, 137],
        [1, 2, 6, 7, 8, 9, 14, 15, 16, 21, 29, 30, 31, 35, 36, 40, 41, 42, 45, 47, 48, 54, 59, 60, 61, 68, 74, 75, 76,
         86,
         87, 92, 93, 94, 110, 111, 112, 115, 116, 128, 129, 132, 137],
        [1, 2, 6, 14, 15, 16, 21, 36, 40, 41, 42, 45, 46, 47, 48, 59, 60, 61, 71, 72, 73, 74, 77, 78, 79, 86, 87, 92,
         103,
         106, 108, 110, 111, 112, 120, 121, 123, 124, 129, 132, 137],
        [6, 7, 8, 9, 14, 15, 16, 20, 24, 29, 30, 31, 35, 36, 40, 48, 54, 59, 60, 61, 68, 74, 86, 87, 92, 93, 94, 98,
         107,
         109, 113, 114, 115, 116, 120, 121, 123, 124, 127, 130, 131],
        [1, 2, 4, 9, 11, 14, 15, 16, 20, 24, 26, 27, 28, 29, 32, 40, 45, 47, 48, 54, 59, 68, 74, 75, 76, 86, 87, 91, 96,
         102, 103, 107, 109, 116, 118, 120, 121, 124, 129, 132, 137],
        [1, 2, 6, 7, 8, 9, 14, 15, 16, 24, 29, 30, 31, 35, 36, 40, 52, 55, 59, 68, 74, 75, 76, 86, 87, 93, 94, 103, 109,
         114, 115, 116, 120, 121, 123, 124, 125, 126, 130, 131, 136],
        [1, 2, 6, 8, 9, 14, 20, 24, 29, 30, 31, 35, 36, 40, 43, 44, 49, 50, 52, 59, 64, 65, 68, 74, 75, 76, 84, 85, 94,
         103, 107, 109, 113, 114, 116, 119, 122, 125, 131, 132, 137],
        [1, 2, 6, 8, 9, 17, 18, 19, 20, 21, 29, 30, 31, 35, 36, 37, 38, 39, 42, 49, 50, 52, 56, 57, 58, 68, 74, 75, 76,
         84, 85, 94, 99, 111, 112, 116, 124, 125, 131, 132, 137],
        [1, 2, 6, 10, 11, 17, 18, 19, 20, 21, 32, 33, 34, 35, 36, 40, 41, 42, 45, 47, 48, 59, 60, 61, 71, 72, 73, 74,
         75,
         76, 86, 87, 88, 95, 96, 103, 106, 108, 117, 132, 137],
        [1, 2, 4, 5, 9, 11, 14, 15, 16, 24, 26, 27, 28, 29, 32, 40, 45, 46, 47, 48, 54, 59, 68, 74, 79, 86, 87, 91, 96,
         102, 103, 107, 109, 116, 118, 120, 121, 124, 129, 136],
        [2, 6, 7, 8, 9, 14, 16, 20, 24, 29, 30, 31, 35, 36, 40, 48, 54, 59, 60, 61, 68, 74, 86, 87, 92, 93, 94, 97, 98,
         107, 109, 113, 114, 115, 116, 120, 123, 124, 130, 131],
        [1, 2, 6, 10, 11, 14, 15, 16, 21, 32, 33, 34, 35, 36, 40, 41, 42, 45, 47, 48, 59, 60, 61, 71, 72, 73, 74, 79,
         86,
         87, 88, 89, 95, 96, 103, 106, 108, 117, 132, 137],
        [1, 2, 6, 10, 11, 12, 13, 17, 32, 33, 34, 35, 36, 40, 42, 45, 46, 47, 48, 51, 59, 61, 71, 72, 73, 74, 77, 78,
         79,
         87, 88, 89, 95, 96, 103, 106, 108, 117, 119, 122],
        [1, 2, 6, 8, 9, 14, 24, 29, 30, 31, 35, 36, 40, 43, 44, 49, 50, 52, 56, 57, 58, 68, 74, 79, 84, 85, 99, 103,
         107,
         109, 113, 114, 116, 119, 122, 125, 126, 133, 134],
        [6, 7, 8, 9, 14, 15, 16, 20, 24, 29, 30, 31, 35, 36, 40, 48, 54, 59, 66, 67, 68, 74, 86, 87, 93, 94, 107, 109,
         113, 114, 115, 116, 120, 121, 123, 124, 130, 131],
        [1, 2, 6, 9, 14, 21, 22, 23, 24, 25, 29, 35, 36, 37, 38, 39, 42, 50, 52, 56, 57, 58, 66, 67, 69, 70, 71, 75, 76,
         84, 85, 112, 116, 124, 125, 126],
        [1, 2, 6, 10, 11, 14, 15, 16, 21, 32, 34, 35, 36, 40, 41, 42, 52, 53, 59, 71, 73, 74, 78, 79, 86, 87, 88, 89,
         96,
         101, 103, 120, 121, 124, 135],
        [1, 2, 6, 14, 21, 36, 37, 38, 39, 42, 49, 50, 52, 56, 57, 58, 71, 72, 73, 74, 77, 78, 79, 84, 85, 99, 104, 112,
         119, 122, 125, 126, 127, 133],
        [1, 2, 4, 5, 9, 11, 12, 13, 17, 24, 26, 27, 28, 29, 32, 40, 45, 46, 47, 48, 51, 54, 59, 68, 74, 79, 88, 89, 96,
         100, 105, 116, 118, 119, 122],
        [1, 2, 6, 9, 14, 29, 35, 36, 40, 44, 59, 64, 65, 66, 67, 69, 70, 71, 75, 76, 80, 81, 84, 85, 104, 114, 116, 119,
         122, 125, 126],
        [1, 2, 6, 9, 14, 20, 24, 29, 30, 35, 36, 40, 44, 59, 64, 65, 68, 74, 75, 76, 80, 81, 82, 84, 85, 104, 114, 116,
         119, 122, 125],
        [1, 2, 6, 14, 16, 21, 36, 37, 38, 39, 42, 52, 53, 59, 62, 63, 71, 73, 74, 84, 85, 92, 101, 103, 111, 112, 119,
         122, 129, 135],
        [1, 2, 6, 10, 11, 14, 15, 16, 21, 32, 35, 36, 40, 41, 42, 52, 53, 59, 71, 74, 79, 86, 87, 90, 96, 103, 120, 121,
         124, 135],
        [1, 2, 6, 9, 14, 24, 29, 30, 35, 36, 40, 44, 59, 64, 65, 68, 74, 75, 76, 83, 84, 85, 104, 114, 116, 119, 122,
         125,
         126],
        [3, 21, 22, 23, 24, 25, 35, 39, 40, 52, 58, 59, 66, 67, 69, 70, 71, 79, 94, 129]
    ]

    print("=" * 70)
    print("baseline PSO")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("Program completed")


if __name__ == "__main__":
    main()