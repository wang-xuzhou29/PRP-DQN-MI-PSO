
import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
import numpy as np
import random
import time
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
from datetime import datetime

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# ===  ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'ACTION_DIM': 3,
    'MIN_VALUES': np.array([1, 1, 1]),
    'MAX_VALUES': np.array([128, 200, 255]),
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 3,
    'PPO_BATCH_SIZE': 64,
    'PPO_EPOCHS': 4,
    'CLIP_EPSILON': 0.2,
    'GAMMA': 0.99,
    'GAE_LAMBDA': 0.95,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'ACTOR_LR': 3e-4,
    'CRITIC_LR': 3e-4,
    'NUM_RUNS': 20,
    'TOP_K_SAMPLES': 20,
    'TARGET_PATHS': [
        [1, 2, 6, 7, 8, 9, 14, 15, 16, 20, 24, 29, 30, 31, 35, 36, 40, 45, 47, 48, 54, 59, 60, 61, 68, 74, 75, 76, 86,
         87, 92, 93, 94, 103, 107, 109, 113, 114, 115, 116, 120, 121, 123, 124, 125, 127, 130, 131, 132, 137],
        [1, 2, 6, 7, 8, 9, 14, 16, 20, 24, 29, 30, 31, 35, 36, 40, 43, 44, 45, 47, 48, 54, 56, 57, 58, 68, 74, 75, 76,
         84,
         85, 92, 93, 94, 103, 107, 109, 113, 114, 115, 116, 119, 122, 125, 130, 131, 132, 137],
        [1, 2, 6, 7, 8, 9, 14, 16, 20, 24, 29, 30, 31, 35, 36, 40, 43, 44, 45, 48, 54, 59, 64, 65, 68, 74, 76, 84, 85,
         93,
         94, 102, 103, 107, 109, 113, 114, 115, 116, 119, 122, 125, 131, 136],
        [1, 2, 6, 7, 8, 9, 17, 18, 19, 20, 21, 29, 30, 31, 35, 36, 37, 38, 39, 42, 45, 47, 48, 54, 56, 57, 58, 68, 74,
         75,
         76, 84, 85, 92, 93, 94, 111, 112, 115, 116, 124, 125, 131, 132, 137],
        [1, 2, 6, 7, 8, 9, 14, 15, 16, 21, 29, 30, 31, 35, 36, 40, 41, 42, 45, 47, 48, 54, 59, 60, 61, 68, 74, 75, 76,
         86,
         87, 92, 93, 94, 110, 111, 112, 115, 116, 128, 129, 132, 137],
        [1, 2, 6, 14, 15, 16, 21, 36, 40, 41, 42, 45, 46, 47, 48, 59, 60, 61, 71, 72, 73, 74, 77, 78, 79, 86, 87, 92,
         103,
         106, 108, 110, 111, 112, 120, 121, 123, 124, 129, 132, 137],
        [6, 7, 8, 9, 14, 15, 16, 20, 24, 29, 30, 31, 35, 36, 40, 48, 54, 59, 60, 61, 68, 74, 86, 87, 92, 93, 94, 98,
         107,
         109, 113, 114, 115, 116, 120, 121, 123, 124, 127, 130, 131],
        [1, 2, 4, 9, 11, 14, 15, 16, 20, 24, 26, 27, 28, 29, 32, 40, 45, 47, 48, 54, 59, 68, 74, 75, 76, 86, 87, 91, 96,
         102, 103, 107, 109, 116, 118, 120, 121, 124, 129, 132, 137],
        [1, 2, 6, 7, 8, 9, 14, 15, 16, 24, 29, 30, 31, 35, 36, 40, 52, 55, 59, 68, 74, 75, 76, 86, 87, 93, 94, 103, 109,
         114, 115, 116, 120, 121, 123, 124, 125, 126, 130, 131, 136],
        [1, 2, 6, 8, 9, 14, 20, 24, 29, 30, 31, 35, 36, 40, 43, 44, 49, 50, 52, 59, 64, 65, 68, 74, 75, 76, 84, 85, 94,
         103, 107, 109, 113, 114, 116, 119, 122, 125, 131, 132, 137],
        [1, 2, 6, 8, 9, 17, 18, 19, 20, 21, 29, 30, 31, 35, 36, 37, 38, 39, 42, 49, 50, 52, 56, 57, 58, 68, 74, 75, 76,
         84, 85, 94, 99, 111, 112, 116, 124, 125, 131, 132, 137],
        [1, 2, 6, 10, 11, 17, 18, 19, 20, 21, 32, 33, 34, 35, 36, 40, 41, 42, 45, 47, 48, 59, 60, 61, 71, 72, 73, 74,
         75,
         76, 86, 87, 88, 95, 96, 103, 106, 108, 117, 132, 137],
        [1, 2, 4, 5, 9, 11, 14, 15, 16, 24, 26, 27, 28, 29, 32, 40, 45, 46, 47, 48, 54, 59, 68, 74, 79, 86, 87, 91, 96,
         102, 103, 107, 109, 116, 118, 120, 121, 124, 129, 136],
        [2, 6, 7, 8, 9, 14, 16, 20, 24, 29, 30, 31, 35, 36, 40, 48, 54, 59, 60, 61, 68, 74, 86, 87, 92, 93, 94, 97, 98,
         107, 109, 113, 114, 115, 116, 120, 123, 124, 130, 131],
        [1, 2, 6, 10, 11, 14, 15, 16, 21, 32, 33, 34, 35, 36, 40, 41, 42, 45, 47, 48, 59, 60, 61, 71, 72, 73, 74, 79,
         86,
         87, 88, 89, 95, 96, 103, 106, 108, 117, 132, 137],
        [1, 2, 6, 10, 11, 12, 13, 17, 32, 33, 34, 35, 36, 40, 42, 45, 46, 47, 48, 51, 59, 61, 71, 72, 73, 74, 77, 78,
         79,
         87, 88, 89, 95, 96, 103, 106, 108, 117, 119, 122],
        [1, 2, 6, 8, 9, 14, 24, 29, 30, 31, 35, 36, 40, 43, 44, 49, 50, 52, 56, 57, 58, 68, 74, 79, 84, 85, 99, 103,
         107,
         109, 113, 114, 116, 119, 122, 125, 126, 133, 134],
        [6, 7, 8, 9, 14, 15, 16, 20, 24, 29, 30, 31, 35, 36, 40, 48, 54, 59, 66, 67, 68, 74, 86, 87, 93, 94, 107, 109,
         113, 114, 115, 116, 120, 121, 123, 124, 130, 131],
        [1, 2, 6, 9, 14, 21, 22, 23, 24, 25, 29, 35, 36, 37, 38, 39, 42, 50, 52, 56, 57, 58, 66, 67, 69, 70, 71, 75, 76,
         84, 85, 112, 116, 124, 125, 126],
        [1, 2, 6, 10, 11, 14, 15, 16, 21, 32, 34, 35, 36, 40, 41, 42, 52, 53, 59, 71, 73, 74, 78, 79, 86, 87, 88, 89,
         96,
         101, 103, 120, 121, 124, 135],
        [1, 2, 6, 14, 21, 36, 37, 38, 39, 42, 49, 50, 52, 56, 57, 58, 71, 72, 73, 74, 77, 78, 79, 84, 85, 99, 104, 112,
         119, 122, 125, 126, 127, 133],
        [1, 2, 4, 5, 9, 11, 12, 13, 17, 24, 26, 27, 28, 29, 32, 40, 45, 46, 47, 48, 51, 54, 59, 68, 74, 79, 88, 89, 96,
         100, 105, 116, 118, 119, 122],
        [1, 2, 6, 9, 14, 29, 35, 36, 40, 44, 59, 64, 65, 66, 67, 69, 70, 71, 75, 76, 80, 81, 84, 85, 104, 114, 116, 119,
         122, 125, 126],
        [1, 2, 6, 9, 14, 20, 24, 29, 30, 35, 36, 40, 44, 59, 64, 65, 68, 74, 75, 76, 80, 81, 82, 84, 85, 104, 114, 116,
         119, 122, 125],
        [1, 2, 6, 14, 16, 21, 36, 37, 38, 39, 42, 52, 53, 59, 62, 63, 71, 73, 74, 84, 85, 92, 101, 103, 111, 112, 119,
         122, 129, 135],
        [1, 2, 6, 10, 11, 14, 15, 16, 21, 32, 35, 36, 40, 41, 42, 52, 53, 59, 71, 74, 79, 86, 87, 90, 96, 103, 120, 121,
         124, 135],
        [1, 2, 6, 9, 14, 24, 29, 30, 35, 36, 40, 44, 59, 64, 65, 68, 74, 75, 76, 83, 84, 85, 104, 114, 116, 119, 122,
         125,
         126],
        [3, 21, 22, 23, 24, 25, 35, 39, 40, 52, 58, 59, 66, 67, 69, 70, 71, 79, 94, 129]
    ],
}
def safe_divide(a, b):
    """安全除法，避免除零错误"""
    if b == 0:
        return 0
    return a / b
# ===  ===
def execute_Tr(moisture, co2, temp):
    """检测低CO2极端条件下的分支触发情况"""
    triggered = set()

    # Fixed: proper indentation and using triggered.add()
    if (co2 < 1150) != (co2 < 1000):
        triggered.add(1)
    if (co2 < 1150) != (co2 < 950):
        triggered.add(2)
    if (co2 < 1150) != (co2 < 1400):
        triggered.add(3)

    # 分支4-11: 调整到中间值区域
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1050 and moisture > 45 and temp > 20):
        triggered.add(4)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1100 and moisture > 45 and temp > 20):
        triggered.add(5)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 or moisture > 45 and temp > 20):
        triggered.add(6)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 35 and temp > 20):
        triggered.add(7)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 30 and temp > 20):
        triggered.add(8)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture < 45 and temp > 20):
        triggered.add(9)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 45 and temp > 5):
        triggered.add(10)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 45 and temp < 20):
        triggered.add(11)

    # 分支12-21: 调整到中间值
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 970 and moisture > 50) or (co2 < 1100 and temp < 22)):
        triggered.add(12)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1000 and moisture > 50) or (co2 < 1100 and temp < 22)):
        triggered.add(13)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 or moisture > 50) or (co2 < 1100 and temp < 22)):
        triggered.add(14)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 40) or (co2 < 1100 and temp < 22)):
        triggered.add(15)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 35) or (co2 < 1100 and temp < 22)):
        triggered.add(16)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 50) and (co2 < 1100 and temp < 22)):
        triggered.add(17)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 50) or (co2 < 960 and temp < 22)):
        triggered.add(18)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 18)):
        triggered.add(19)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 50) or (co2 < 1100 and temp > 22)):
        triggered.add(20)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 50) or (co2 > 1100 and temp < 22)):
        triggered.add(21)

    # 分支22-25: 调整阈值到中间 - using safe_divide
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp < 22) != (safe_divide(moisture, co2 - 700) > 0.04 and temp < 22):
        triggered.add(22)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp < 22) != (safe_divide(moisture, co2 - 700) > 0.03 and temp < 22):
        triggered.add(23)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp < 22) != (safe_divide(moisture, co2 - 700) > 0.06 or temp < 22):
        triggered.add(24)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp < 22) != (moisture + safe_divide(100, co2 - 700) > 0.06 and temp < 22):
        triggered.add(25)

    # 分支26-36: 调整到中间值
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1000 and moisture > 45 and temp > 20):
        triggered.add(26)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1070 and moisture > 45 and temp > 20):
        triggered.add(27)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 > 1150 and moisture > 45 and temp > 20):
        triggered.add(28)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture < 45 and temp > 20):
        triggered.add(29)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 25 and temp > 20):
        triggered.add(30)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 30 and temp > 20):
        triggered.add(31)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 45 and temp < 20):
        triggered.add(32)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 45 and temp > 15):
        triggered.add(33)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 45 and temp > 12):
        triggered.add(34)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 45 or temp > 20):
        triggered.add(35)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 or moisture > 45 and temp > 20):
        triggered.add(36)

    # 分支37-44: 调整到中间值
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 970 and moisture < 40 and temp < 22):
        triggered.add(37)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 1020 and moisture < 40 and temp < 22):
        triggered.add(38)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 > 1150 and moisture < 40 and temp < 22):
        triggered.add(39)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 1150 or moisture < 40 and temp < 22):
        triggered.add(40)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 1150 and moisture < 50 and temp < 22):
        triggered.add(41)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 1150 and moisture > 40 and temp < 22):
        triggered.add(42)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 1150 and moisture < 40 and temp < 27):
        triggered.add(43)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 1150 and moisture < 40 and temp < 32):
        triggered.add(44)

    # 分支45-55: 扩大范围
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1050 and 35 < moisture < 60 and 15 < temp < 28):
        triggered.add(45)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1120 and 35 < moisture < 60 and 15 < temp < 28):
        triggered.add(46)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1070 and 35 < moisture < 60 and 15 < temp < 28):
        triggered.add(47)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 > 1150 and 35 < moisture < 60 and 15 < temp < 28):
        triggered.add(48)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 25 < moisture < 60 and 15 < temp < 28):
        triggered.add(49)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 20 < moisture < 60 and 15 < temp < 28):
        triggered.add(50)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 35 < moisture < 52 and 15 < temp < 28):
        triggered.add(51)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 35 < moisture < 60 or 15 < temp < 28):
        triggered.add(52)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 35 < moisture < 60 and 10 < temp < 28):
        triggered.add(53)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 35 < moisture < 60 and 15 < temp < 20):
        triggered.add(54)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 35 < moisture < 60 and 15 < temp < 33):
        triggered.add(55)

    # 分支56-65: 调整到中间值和扩大范围
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1020 and moisture < 40 and 15 < temp < 25):
        triggered.add(56)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1070 and moisture < 40 and 15 < temp < 25):
        triggered.add(57)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 > 1150 and moisture < 40 and 15 < temp < 25):
        triggered.add(58)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 or moisture < 40 and 15 < temp < 25):
        triggered.add(59)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 and moisture < 50 and 15 < temp < 25):
        triggered.add(60)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 and moisture < 55 and 15 < temp < 25):
        triggered.add(61)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 and moisture < 40 and 10 < temp < 25):
        triggered.add(62)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 and moisture < 40 and 5 < temp < 25):
        triggered.add(63)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 and moisture < 40 and 15 < temp < 30):
        triggered.add(64)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 and moisture < 40 and 15 < temp < 30):
        triggered.add(65)

    # 分支66-74: 调整比例和温度阈值
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 850) > 0.06 and temp > 20):
        triggered.add(66)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 900) > 0.06 and temp > 20):
        triggered.add(67)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 + 200) > 0.06 and temp > 20):
        triggered.add(68)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 700) > 0.04 and temp > 20):
        triggered.add(69)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 700) > 0.03 and temp > 20):
        triggered.add(70)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 700) > 0.06 or temp > 20):
        triggered.add(71)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 700) > 0.06 and temp > 15):
        triggered.add(72)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 700) > 0.06 and temp > 12):
        triggered.add(73)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 700) > 0.06 and temp < 20):
        triggered.add(74)

    if (co2 + moisture > 1150 and temp > 18) != (co2 + moisture > 1100 and temp > 18):
        triggered.add(75)
    if (co2 + moisture > 1150 and temp > 18) != (co2 + moisture > 1050 and temp > 18):
        triggered.add(76)
    if (co2 + moisture > 1150 and temp > 18) != (co2 + moisture > 1150 and temp > 15):
        triggered.add(77)
    if (co2 + moisture > 1150 and temp > 18) != (co2 + moisture > 1150 and temp > 12):
        triggered.add(78)
    if (co2 + moisture > 1150 and temp > 18) != (co2 + moisture > 1150 and temp < 18):
        triggered.add(79)

    if (co2 < 1150 and abs(moisture - 45) > 15 and abs(temp - 20) > 7) != (
            co2 < 1000 and abs(moisture - 45) > 15 and abs(temp - 20) > 7):
        triggered.add(80)
    if (co2 < 1150 and abs(moisture - 45) > 15 and abs(temp - 20) > 7) != (
            co2 < 1050 and abs(moisture - 45) > 15 and abs(temp - 20) > 7):
        triggered.add(81)
    if (co2 < 1150 and abs(moisture - 45) > 15 and abs(temp - 20) > 7) != (
            co2 < 1150 and abs(moisture - 42) > 15 and abs(temp - 20) > 7):
        triggered.add(82)
    if (co2 < 1150 and abs(moisture - 45) > 15 and abs(temp - 20) > 7) != (
            co2 < 1150 and abs(moisture - 45) > 12 and abs(temp - 20) > 7):
        triggered.add(83)

    # 分支84-87: 简化条件到中间值
    if (co2 < 1150 and moisture < 40) != (co2 < 1000 and moisture < 40):
        triggered.add(84)
    if (co2 < 1150 and moisture < 40) != (co2 < 1050 and moisture < 40):
        triggered.add(85)
    if (co2 < 1150 and moisture < 40) != (co2 < 1150 and moisture < 50):
        triggered.add(86)
    if (co2 < 1150 and moisture < 40) != (co2 < 1150 and moisture < 55):
        triggered.add(87)

    # 分支88-92: 调整到中间范围
    if (co2 < 1150 and 12 < temp < 25 and moisture > 45) != (co2 < 1030 and 12 < temp < 25 and moisture > 45):
        triggered.add(88)
    if (co2 < 1150 and 12 < temp < 25 and moisture > 45) != (co2 < 1080 and 12 < temp < 25 and moisture > 45):
        triggered.add(89)
    if (co2 < 1150 and 12 < temp < 25 and moisture > 45) != (co2 < 1150 and 8 < temp < 25 and moisture > 45):
        triggered.add(90)
    if (co2 < 1150 and 12 < temp < 25 and moisture > 45) != (co2 < 1150 and 12 < temp < 30 and moisture > 45):
        triggered.add(91)
    if (co2 < 1150 and 12 < temp < 25 and moisture > 45) != (co2 < 1150 and 12 < temp < 25 and moisture > 35):
        triggered.add(92)

    # 分支93-96: 调整到中间值
    if (moisture > 45 and temp > 20) != (moisture > 35 and temp > 20):
        triggered.add(93)
    if (moisture > 45 and temp > 20) != (moisture > 32 and temp > 20):
        triggered.add(94)
    if (moisture > 45 and temp > 20) != (moisture > 45 and temp > 15):
        triggered.add(95)
    if (moisture > 45 and temp > 20) != (moisture > 45 and temp < 20):
        triggered.add(96)

    # 分支97-102: 扩大范围
    if (1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 25) != (
            950 < co2 < 1250 and 35 < moisture < 55 and 15 < temp < 25):
        triggered.add(97)
    if (1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 25) != (
            900 < co2 < 1300 and 35 < moisture < 55 and 15 < temp < 25):
        triggered.add(98)
    if (1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 25) != (
            1000 < co2 < 1200 and 28 < moisture < 55 and 15 < temp < 25):
        triggered.add(99)
    if (1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 25) != (
            1000 < co2 < 1200 and 35 < moisture < 62 and 15 < temp < 25):
        triggered.add(100)
    if (1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 25) != (
            1000 < co2 < 1200 and 35 < moisture < 55 and 12 < temp < 25):
        triggered.add(101)
    if (1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 25) != (
            1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 28):
        triggered.add(102)

    # 分支103-109: 调整条件和扩大范围
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1050 and 30 < moisture < 55 and (temp < 20 or temp > 22)):
        triggered.add(103)
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1150 and 22 < moisture < 55 and (temp < 20 or temp > 22)):
        triggered.add(104)
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1150 and 30 < moisture < 62 and (temp < 20 or temp > 22)):
        triggered.add(105)
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1150 and 30 < moisture < 55 and (temp < 16 or temp > 22)):
        triggered.add(106)
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 26)):
        triggered.add(107)
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1150 and 30 < moisture < 55 and (temp < 15 or temp > 22)):
        triggered.add(108)
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 28)):
        triggered.add(109)

    # 分支110-114: 调整到中间值
    if (moisture < 45 and co2 < 1150 and temp < 22) != (moisture < 38 and co2 < 1150 and temp < 22):
        triggered.add(110)
    if (moisture < 45 and co2 < 1150 and temp < 22) != (moisture < 35 and co2 < 1150 and temp < 22):
        triggered.add(111)
    if (moisture < 45 and co2 < 1150 and temp < 22) != (moisture < 45 and co2 < 1000 and temp < 22):
        triggered.add(112)
    if (moisture < 45 and co2 < 1150 and temp < 22) != (moisture < 45 and co2 < 1150 and temp < 27):
        triggered.add(113)
    if (moisture < 45 and co2 < 1150 and temp < 22) != (moisture < 45 and co2 < 1150 and temp < 30):
        triggered.add(114)

    # 分支115-118: 调整条件
    if (co2 < 1200 and moisture > 45 and temp > 20 and co2 < 1150) != (
            co2 < 1200 and moisture > 35 and temp > 20 and co2 < 1150):
        triggered.add(115)
    if (co2 < 1200 and moisture > 45 and temp > 20 and co2 < 1150) != (
            co2 < 1200 and moisture < 45 and temp > 20 and co2 < 1150):
        triggered.add(116)
    if (co2 < 1200 and moisture > 45 and temp > 20 and co2 < 1150) != (
            co2 < 1200 and moisture > 45 and temp > 15 and co2 < 1150):
        triggered.add(117)
    if (co2 < 1200 and moisture > 45 and temp > 20 and co2 < 1150) != (
            co2 < 1200 and moisture > 45 and temp > 20 and co2 < 1070):
        triggered.add(118)

    # 分支119-124: 调整到中间值和扩大范围
    if (co2 < 1150 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)) != (
            co2 < 1000 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)):
        triggered.add(119)
    if (co2 < 1150 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)) != (
            co2 < 1150 and (moisture > 30 or moisture < 40) and (temp > 22 or temp < 18)):
        triggered.add(120)
    if (co2 < 1150 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)) != (
            co2 < 1150 and (moisture > 40 or moisture < 40) and (temp > 22 or temp < 18)):
        triggered.add(121)
    if (co2 < 1150 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)) != (
            co2 < 1150 and (moisture > 50 and moisture < 40) and (temp > 22 or temp < 18)):
        triggered.add(122)
    if (co2 < 1150 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)) != (
            co2 < 1150 and (moisture > 50 or moisture < 45) and (temp > 22 or temp < 18)):
        triggered.add(123)
    if (co2 < 1150 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)) != (
            co2 < 1150 and (moisture > 50 or moisture < 40) or (temp > 22 or temp < 18)):
        triggered.add(124)

    # 分支125-131: 调整比例和阈值
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1050 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45):
        triggered.add(125)
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1100 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45):
        triggered.add(126)
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1150 and safe_divide(temp, moisture + 5) > 0.5 and moisture < 45):
        triggered.add(127)
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1150 and safe_divide(temp, moisture + 1) > 0.4 and moisture < 45):
        triggered.add(128)
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 or moisture < 45):
        triggered.add(129)
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 38):
        triggered.add(130)
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 35):
        triggered.add(131)

    # 分支132-137: 调整范围到中间
    if (co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 25) != (
            co2 < 1000 and 35 <= moisture <= 50 and 15 <= temp <= 25):
        triggered.add(132)
    if (co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 25) != (
            co2 < 1150 and 28 <= moisture <= 50 and 15 <= temp <= 25):
        triggered.add(133)
    if (co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 25) != (
            co2 < 1150 and 32 <= moisture <= 50 and 15 <= temp <= 25):
        triggered.add(134)
    if (co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 25) != (
            co2 < 1150 and 35 <= moisture <= 50 and 12 <= temp <= 25):
        triggered.add(135)
    if (co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 25) != (
            co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 28):
        triggered.add(136)
    if (co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 25) != (
            co2 < 1050 and 35 <= moisture <= 50 and 15 <= temp <= 25):
        triggered.add(137)

    return triggered
# === 状态处理辅助函数 ===
def clip_state(state):
    return np.clip(state, EXPERIMENT_CONFIG['MIN_VALUES'], EXPERIMENT_CONFIG['MAX_VALUES'])

def denormalize_state(normalized_state):
    """将归一化状态还原为原始状态"""
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return normalized_state * (max_vals - min_vals) / 2 + (min_vals + max_vals) / 2

def coverage_similarity(triggered, target_path):
    """
    Similarity:  / target paths
    
    """
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0

    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)

def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']

    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']

    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']

    return reward

# === PPO Actor ===
class PPOActor(nn.Module):
    def __init__(self, state_dim=3, action_dim=3, hidden_dim=256):
        super(PPOActor, self).__init__()

        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)

        self.mean_head = nn.Linear(hidden_dim, action_dim)
        self.log_std_head = nn.Linear(hidden_dim, action_dim)

        self.action_scale = 10.0

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))
        x = F.relu(self.fc3(x))

        mean = torch.tanh(self.mean_head(x)) * self.action_scale
        log_std = torch.clamp(self.log_std_head(x), -20, 2)

        return mean, log_std

    def get_action_and_log_prob(self, state):
        mean, log_std = self.forward(state)
        std = torch.exp(log_std)

        dist = torch.distributions.Normal(mean, std)
        action = dist.sample()
        log_prob = dist.log_prob(action).sum(dim=-1)

        return action, log_prob

    def get_log_prob(self, state, action):
        mean, log_std = self.forward(state)
        std = torch.exp(log_std)

        dist = torch.distributions.Normal(mean, std)
        log_prob = dist.log_prob(action).sum(dim=-1)

        return log_prob

# === PPO Critic ===
class PPOCritic(nn.Module):
    def __init__(self, state_dim=3, hidden_dim=256):
        super(PPOCritic, self).__init__()

        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)
        self.value_head = nn.Linear(hidden_dim, 1)

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))
        x = F.relu(self.fc3(x))
        value = self.value_head(x)
        return value.squeeze(-1)

# === PPO ===
class PPOBuffer:
    def __init__(self):
        self.states = []
        self.actions = []
        self.rewards = []
        self.values = []
        self.log_probs = []
        self.dones = []
        self.path_indices = []
        self.similarities = []

    def store(self, state, action, reward, value, log_prob, done, path_idx, similarity):
        self.states.append(state)
        self.actions.append(action)
        self.rewards.append(reward)
        self.values.append(value)
        self.log_probs.append(log_prob)
        self.dones.append(done)
        self.path_indices.append(path_idx)
        self.similarities.append(similarity)

    def compute_advantages(self):
        rewards = np.array(self.rewards)
        values = np.array(self.values)
        dones = np.array(self.dones)

        advantages = np.zeros_like(rewards)
        returns = np.zeros_like(rewards)
        last_advantage = 0

        for t in reversed(range(len(rewards))):
            if t == len(rewards) - 1:
                next_value = 0
            else:
                next_value = values[t + 1]

            delta = rewards[t] + EXPERIMENT_CONFIG['GAMMA'] * next_value * (1 - dones[t]) - values[t]
            advantages[t] = delta + EXPERIMENT_CONFIG['GAMMA'] * EXPERIMENT_CONFIG['GAE_LAMBDA'] * (
                    1 - dones[t]) * last_advantage
            last_advantage = advantages[t]

        returns = advantages + values
        advantages = (advantages - np.mean(advantages)) / (np.std(advantages) + 1e-8)

        return advantages, returns

    def get_batch_iterator(self, advantages, returns):
        indices = np.arange(len(self.states))
        np.random.shuffle(indices)

        batch_size = EXPERIMENT_CONFIG['PPO_BATCH_SIZE']
        for start in range(0, len(self.states), batch_size):
            end = min(start + batch_size, len(self.states))
            batch_indices = indices[start:end]

            yield {
                'states': torch.FloatTensor(np.array([self.states[i] for i in batch_indices])).to(device),
                'actions': torch.FloatTensor(np.array([self.actions[i] for i in batch_indices])).to(device),
                'old_log_probs': torch.FloatTensor(np.array([self.log_probs[i] for i in batch_indices])).to(device),
                'advantages': torch.FloatTensor(advantages[batch_indices]).to(device),
                'returns': torch.FloatTensor(returns[batch_indices]).to(device)
            }

    def get_top_k_per_path(self, num_paths, k=20):
        path_samples = {i: [] for i in range(num_paths)}

        for idx in range(len(self.states)):
            path_idx = self.path_indices[idx]
            similarity = self.similarities[idx]
            state = self.states[idx]
            path_samples[path_idx].append((idx, similarity, state))

        top_k_results = {}
        for path_idx in range(num_paths):
            samples = path_samples[path_idx]
            if len(samples) == 0:
                top_k_results[path_idx] = []
                continue

            samples.sort(key=lambda x: x[1], reverse=True)
            top_k = samples[:k]

            top_k_results[path_idx] = []
            for sample in top_k:
                normalized_state = sample[2]
                original_state = denormalize_state(normalized_state)
                original_state_int = np.round(original_state).astype(int)

                top_k_results[path_idx].append({
                    'state': original_state_int,
                    'similarity': sample[1],
                    'triggered': execute_Tr(*original_state_int)
                })

        return top_k_results

    def clear(self):
        self.states.clear()
        self.actions.clear()
        self.rewards.clear()
        self.values.clear()
        self.log_probs.clear()
        self.dones.clear()
        self.path_indices.clear()
        self.similarities.clear()

    def __len__(self):
        return len(self.states)

# === PPO ===
class PPOAgent:
    def __init__(self, state_dim=3, action_dim=3):
        self.state_dim = state_dim
        self.action_dim = action_dim

        self.actor = PPOActor(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.actor_optimizer = optim.Adam(self.actor.parameters(), lr=EXPERIMENT_CONFIG['ACTOR_LR'])

        self.critic = PPOCritic(state_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_optimizer = optim.Adam(self.critic.parameters(), lr=EXPERIMENT_CONFIG['CRITIC_LR'])

        self.buffer = PPOBuffer()
        self.update_count = 0

    def get_action(self, state):
            min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
            max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
            normalized_state = (state - (min_vals + max_vals) / 2) / ((max_vals - min_vals) / 2)
            state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)

            with torch.no_grad():
                action, log_prob = self.actor.get_action_and_log_prob(state_tensor)
                value = self.critic(state_tensor)

            action = action.cpu().numpy()[0]
            log_prob = log_prob.cpu().item()
            value = value.cpu().item()

            return action, log_prob, value
    def store_experience(self, state, action, reward, value, log_prob, done, path_idx, similarity):
            min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
            max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
            normalized_state = (state - (min_vals + max_vals) / 2) / ((max_vals - min_vals) / 2)
            self.buffer.store(normalized_state, action, reward, value, log_prob, done, path_idx, similarity)

    def update(self):
        if len(self.buffer) == 0:
            return

        advantages, returns = self.buffer.compute_advantages()

        for epoch in range(EXPERIMENT_CONFIG['PPO_EPOCHS']):
            for batch in self.buffer.get_batch_iterator(advantages, returns):
                new_log_probs = self.actor.get_log_prob(batch['states'], batch['actions'])
                ratio = torch.exp(new_log_probs - batch['old_log_probs'])

                surr1 = ratio * batch['advantages']
                surr2 = torch.clamp(ratio, 1 - EXPERIMENT_CONFIG['CLIP_EPSILON'],
                                    1 + EXPERIMENT_CONFIG['CLIP_EPSILON']) * batch['advantages']
                actor_loss = -torch.min(surr1, surr2).mean()

                self.actor_optimizer.zero_grad()
                actor_loss.backward()
                torch.nn.utils.clip_grad_norm_(self.actor.parameters(), 0.5)
                self.actor_optimizer.step()

                new_values = self.critic(batch['states'])
                critic_loss = F.mse_loss(new_values, batch['returns'])

                self.critic_optimizer.zero_grad()
                critic_loss.backward()
                torch.nn.utils.clip_grad_norm_(self.critic.parameters(), 0.5)
                self.critic_optimizer.step()

        self.update_count += 1
        self.buffer.clear()

        if self.update_count % 2 == 0:
            print(f"  -> PPOcompleted (Run {self.update_count})")

# === Metric ===
def calculate_run_performance(run_idx, ppo_results, training_time, total_steps, update_count, agent):
    """ runMetric"""
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    # 1. (Total Reward)
    total_reward = 0
    # 2. (Average Reward)
    average_reward = 0
    # 5. (Convergence)
    convergence = 0
    # 12. (Environment Adaptability)
    environment_adaptability = 0
    # 13. (Generalization Ability)
    generalization_ability = 0
    # 15. (Computational Efficiency)
    computational_efficiency = 0
    # 16. (Policy Update Frequency)
    policy_update_frequency = 0

    # Similarity
    all_similarities = []

    # Metric
    total_samples = 0
    all_rewards = []

    for path_idx in range(num_paths):
        samples = ppo_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']

            total_reward += reward
            all_rewards.append(reward)
            all_similarities.append(similarity)
            total_samples += 1

    # 1. 
    total_reward = total_reward

    # 2. 
    if total_samples > 0:
        average_reward = total_reward / total_samples

    # 5. (Average Similarity)
    if all_similarities:
        convergence = np.mean(all_similarities)

    # 12. (Similarity)
    if len(all_similarities) > 1:
        environment_adaptability = 1 / (np.std(all_similarities) + 1e-8)

    # 13. (Average Similarity)
    generalization_ability = convergence

    # 15. (/ seconds)
    if training_time > 0:
        computational_efficiency = total_steps / training_time

    # 16. 
    if training_time > 0:
        policy_update_frequency = update_count / training_time

    # Similarity
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0

    return {
        'Run': run_idx + 1,

        # Metric
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),

        # Similarity
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }

# === Excel ===
def export_to_excel(all_ppo_results, all_performance_data, target_paths, output_path="PPO_20 run.xlsx"):
    """20 runPPOExcel"""
    print("\nExcel...")

    # 
    all_ppo_summary_data = []
    all_ppo_detailed_data = []

    #  run
    for run_idx, (ppo_results, performance_data) in enumerate(zip(all_ppo_results, all_performance_data)):
        # ===== Sheet1: PPOPath  =====
        ppo_summary_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = ppo_results[path_idx]

            if len(samples) == 0:
                ppo_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue

            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = '' if perfect_count > 0 else ''

            ppo_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })

        all_ppo_summary_data.extend(ppo_summary_data)

        # ===== Sheet2: PPODetailed Sample Data =====
        ppo_detailed_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = ppo_results[path_idx]

            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']

                ppo_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })

        all_ppo_detailed_data.extend(ppo_detailed_data)

    # Excel
    ppo_summary_df = pd.DataFrame(all_ppo_summary_data)
    ppo_detailed_df = pd.DataFrame(all_ppo_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet1: PPOPath 
        ppo_summary_df.to_excel(writer, sheet_name='PPOPath ', index=False)

        # Sheet2: PPODetailed Sample Data
        ppo_detailed_df.to_excel(writer, sheet_name='PPODetailed Sample Data', index=False)

        # Sheet3: Metric - 
        selected_columns = [
            'Run',
            'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df_selected = performance_df[selected_columns]
        performance_df_selected.to_excel(writer, sheet_name='Metric', index=False)

        # 
        workbook = writer.book

        # 
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        perfect_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # 

        # === Sheet1 ===
        ws1 = writer.sheets['PPOPath ']
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 
        for row_idx in range(2, ws1.max_row + 1):
            if ws1.cell(row_idx, 9).value == '':  # Run 9""
                for col_idx in range(1, ws1.max_column + 1):
                    ws1.cell(row_idx, col_idx).fill = perfect_fill

        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 15
        ws1.column_dimensions['F'].width = 15
        ws1.column_dimensions['G'].width = 15
        ws1.column_dimensions['H'].width = 15
        ws1.column_dimensions['I'].width = 15
        ws1.column_dimensions['J'].width = 50

        # === Sheet2 ===
        ws2 = writer.sheets['PPODetailed Sample Data']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 15
        ws2.column_dimensions['I'].width = 40
        ws2.column_dimensions['J'].width = 40
        ws2.column_dimensions['K'].width = 15
        ws2.column_dimensions['L'].width = 15

        # === Sheet3 ===
        ws3 = writer.sheets['Metric']
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        for col in columns:
            ws3.column_dimensions[col].width = 18

    print(f"Excel: {output_path}")
    print(f"  - Sheet1: PPOPath  ({len(all_ppo_summary_data)})")
    print(f"  - Sheet2: PPODetailed Sample Data ({len(all_ppo_detailed_data)})")
    print(f"  - Sheet3: Metric ({len(all_performance_data)})")

# === PPO ===
def train_ppo_workflow():
    print("=" * 80)
    print("PPO")
    print("Similarity:  / target paths")
    print("=" * 80)

    agent = PPOAgent()
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n: Path {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']}")
    path_samples = {}
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']

    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            # 分别生成 X, Y, Z 的随机整数
            state = np.array([
                np.random.randint(min_vals[0], max_vals[0] + 1),
                np.random.randint(min_vals[1], max_vals[1] + 1),
                np.random.randint(min_vals[2], max_vals[2] + 1)
            ], dtype=np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  Path  {path_idx + 1}/{num_paths}:  {len(samples)} ")

    # 
    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size

    print(f"\n: {batch_size},{EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(f": {num_batches} /Path  x {num_paths} Path  = {num_batches * num_paths} ")
    print("-" * 80)

    global_buffer = PPOBuffer()

    for batch_idx in range(num_batches):
        print(f"\n {batch_idx + 1}/{num_batches}")

        for path_idx in range(num_paths):
            target_path = target_paths[path_idx]
            batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]

            batch_rewards = []
            batch_similarities = []

            for sample_idx, initial_state in enumerate(batch_samples):
                state = initial_state.copy()
                episode_reward = 0
                final_similarity = 0

                for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                    action, log_prob, value = agent.get_action(state)

                    next_state = state + action
                    next_state = clip_state(next_state)

                    triggered = execute_Tr(*next_state)
                    reward = unified_reward_function(triggered, target_path)
                    similarity = coverage_similarity(triggered, target_path)

                    done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)

                    agent.store_experience(state, action, reward, value, log_prob, done, path_idx, similarity)
                    normalized_state = (state - (min_vals + max_vals) / 2) / ((max_vals - min_vals) / 2)
                    global_buffer.store(
                        normalized_state,
                        action, reward, value, log_prob, done, path_idx, similarity
                    )

                    state = next_state
                    episode_reward += reward
                    final_similarity = similarity
                    total_steps += 1

                batch_rewards.append(episode_reward)
                batch_similarities.append(final_similarity)

            avg_reward = np.mean(batch_rewards)
            avg_similarity = np.mean(batch_similarities)
            print(f"  Path {path_idx + 1}: ={avg_reward:.2f}, Average Similarity={avg_similarity:.4f}")

        print(f"\n  PPO...")
        agent.update()
        print(f"  : {len(global_buffer)}")

    training_time = time.time() - start_time

    print("\n" + "=" * 80)
    print(f"PPOcompleted! Total elapsed time: {training_time:.2f} seconds, : {total_steps}")
    print(f": {len(global_buffer)}")
    print(f"PPO: {agent.update_count}")
    print("=" * 80)

    # Top-K
    print(f"\nPath SimilarityMaximum{EXPERIMENT_CONFIG['TOP_K_SAMPLES']}...")
    ppo_top_k_results = global_buffer.get_top_k_per_path(num_paths, EXPERIMENT_CONFIG['TOP_K_SAMPLES'])

    return agent, ppo_top_k_results, training_time, total_steps, agent.update_count

# ===  ===
def main():
    print("\n" + "=" * 80)
    print("PPO - 20 run")
    print("Metric")
    print("=" * 80)

    all_ppo_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    # 20
    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'=' * 80}")
        print(f"Start run  {run_idx + 1}/{EXPERIMENT_CONFIG['NUM_RUNS']}  run")
        print(f"{'=' * 80}")

        # PPO
        ppo_agent, ppo_results, training_time, total_steps, update_count = train_ppo_workflow()

        # Metric
        performance_data = calculate_run_performance(
            run_idx, ppo_results, training_time, total_steps, update_count, ppo_agent
        )

        # 
        all_ppo_results.append(ppo_results)
        all_performance_data.append(performance_data)

        print(f"\nRun  {run_idx + 1}  runcompleted!")
        print(f"  Total Reward: {performance_data['Total Reward']}")
        print(f"  Average Reward: {performance_data['Average Reward']}")
        print(f"  Average Similarity: {performance_data['Average Similarity']}")

    # Excel(20 run)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"PPO_20 run_{timestamp}.xlsx"
    export_to_excel(all_ppo_results, all_performance_data, target_paths, output_path)

    # 
    print("\n" + "=" * 80)
    print("20 run")
    print("=" * 80)

    # Metric
    total_rewards = [p['Total Reward'] for p in all_performance_data]
    average_rewards = [p['Average Reward'] for p in all_performance_data]
    convergences = [p['Convergence'] for p in all_performance_data]
    environment_adaptabilities = [p['Environment Adaptability'] for p in all_performance_data]
    generalization_abilities = [p['Generalization Ability'] for p in all_performance_data]
    computational_efficiencies = [p['Computational Efficiency'] for p in all_performance_data]
    policy_update_frequencies = [p['Policy Update Frequency'] for p in all_performance_data]
    avg_similarities = [p['Average Similarity'] for p in all_performance_data]

    print(f"\nTotal Reward Statistics:")
    print(f"  Mean: {np.mean(total_rewards):.2f}")
    print(f"  Standard deviation: {np.std(total_rewards):.2f}")

    print(f"\nAverage Reward Statistics:")
    print(f"  Mean: {np.mean(average_rewards):.4f}")
    print(f"  Standard deviation: {np.std(average_rewards):.4f}")

    print(f"\nConvergence Statistics:")
    print(f"  Mean: {np.mean(convergences):.4f}")
    print(f"  Standard deviation: {np.std(convergences):.4f}")

    print(f"\nEnvironment Adaptability Statistics:")
    print(f"  Mean: {np.mean(environment_adaptabilities):.4f}")
    print(f"  Standard deviation: {np.std(environment_adaptabilities):.4f}")

    print(f"\nGeneralization Ability Statistics:")
    print(f"  Mean: {np.mean(generalization_abilities):.4f}")
    print(f"  Standard deviation: {np.std(generalization_abilities):.4f}")

    print(f"\nComputational Efficiency Statistics:")
    print(f"  Mean: {np.mean(computational_efficiencies):.2f}")
    print(f"  Standard deviation: {np.std(computational_efficiencies):.2f}")

    print(f"\nPolicy Update Frequency Statistics:")
    print(f"  Mean: {np.mean(policy_update_frequencies):.4f}")
    print(f"  Standard deviation: {np.std(policy_update_frequencies):.4f}")

    print("\n" + "=" * 80)
    print(f" {EXPERIMENT_CONFIG['NUM_RUNS']} completed!")
    print("=" * 80)

if __name__ == "__main__":
    main()