import os
import random
from collections import deque
import numpy as np
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

import warnings
warnings.filterwarnings('ignore', 'overflow encountered')

# --- 全局状态范围配置 ---
STATE_MIN_X, STATE_MAX_X = 2, 100
STATE_MIN_Y, STATE_MAX_Y = 1, 105
STATE_MIN_Z, STATE_MAX_Z = 1, 110

def generate_input():
    return [
        random.randint(STATE_MIN_X, STATE_MAX_X),
        random.randint(STATE_MIN_Y, STATE_MAX_Y),
        random.randint(STATE_MIN_Z, STATE_MAX_Z)
    ]

def execute_Tr(x, y, z):
    triggered = set()

    # --- 分支 1-14 ---
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 1) > 120): triggered.add(1)
    if ((y * z) / (x + 1) > 100) != ((y * y) / (x + 1) > 100): triggered.add(2)
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 8) > 100): triggered.add(3)
    if ((y * z) / (x + 1) > 100) != ((y + z) / (x + 1) > 100): triggered.add(4)
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x * 3 + 1) > 100): triggered.add(5)
    if ((y * z) / (x + 1) > 100) != ((z * z) / (x + 1) > 100): triggered.add(6)
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 1) > 80): triggered.add(7)
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + y) > 100): triggered.add(8)
    if ((y * z) / (x + 1) > 100) != ((y * 10) / (x + 1) > 100): triggered.add(9)
    if ((y * z) / (x + 1) > 100) != ((x * z) / (x + 1) > 100): triggered.add(10)
    if ((y * z) / (x + 1) > 100) != ((y * z * 2) / (x + 1) > 100): triggered.add(11)
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 5) > 100): triggered.add(12)
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 1) > 200): triggered.add(13)
    if ((y * z) / (x + 1) > 100) != ((5 * z) / (x + 1) > 100): triggered.add(14)

    # --- 分支 15-22 ---
    if ((z - x) < 0.3 * y) != ((z * 1.2 - x) < 0.3 * y): triggered.add(15)
    if ((z - x) < 0.3 * y) != ((z - x * 1.1) < 0.3 * y): triggered.add(16)
    if ((z - x) < 0.3 * y) != ((z - x) < 0.5 * y): triggered.add(17)
    if ((z - x) < 0.3 * y) != ((z - x) < 0.3 * z): triggered.add(18)
    if ((z - x) < 0.3 * y) != ((z - x) < 0.3 * y * y): triggered.add(19)
    if ((z - x) < 0.3 * y) != ((z - y) < 0.3 * y): triggered.add(20)
    if ((z - x) < 0.3 * y) != ((z - x) < 0.8 * y): triggered.add(21)
    if ((z - x) < 0.3 * y) != ((z - x) < 0.3 * x): triggered.add(22)

    # --- 分支 23-36 ---
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 3): triggered.add(23)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 * y ** 3) < z ** 4): triggered.add(24)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2 + y ** 3) < z ** 2): triggered.add(25)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 2) < z ** 2): triggered.add(26)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * 3) < (z ** 2)): triggered.add(27)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * x ** 3) < z ** 2): triggered.add(28)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 1 + y ** 3) < z ** 2): triggered.add(29)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 1) < z ** 2): triggered.add(30)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x * 3 + y ** 3) < z ** 2): triggered.add(31)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x * 3 + y ** 3) < z ** 2): triggered.add(32)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * 3) < z ** 2): triggered.add(33)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** x + y ** 3) < z ** 2): triggered.add(34)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2 + y ** 3) < z ** 2): triggered.add(35)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 1.5): triggered.add(36)

    # --- 分支 37-54 ---
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 or (y / (z + 0.001) < 0.2)): triggered.add(37)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y % (z + 0.001) < 0.2)): triggered.add(38)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y / (z - 0.001) < 0.4)): triggered.add(39)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.3)): triggered.add(40)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.1)): triggered.add(41)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 8 and (y / (z + 0.001) < 0.2)): triggered.add(42)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 2 and (y / (z + 0.001) < 0.2)): triggered.add(43)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 1 and (y / (z + 0.001) < 0.2)): triggered.add(44)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 10 and (y / (z + 0.001) < 0.2)): triggered.add(45)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x * y / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.2)): triggered.add(46)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x * x / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.2)): triggered.add(47)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y * x / (z + 0.001) < 0.2)): triggered.add(48)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y * y / (z + 0.001) < 0.2)): triggered.add(49)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y * z / (z + 0.001) < 0.2)): triggered.add(50)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.5)): triggered.add(51)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y - (z + 0.001) < 0.2)): triggered.add(52)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x + (y + 0.001)) > 5 and (y / (z + 0.001) < 0.2)): triggered.add(53)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x - (y + 0.001)) > 5 and (y / (z + 0.001) < 0.2)): triggered.add(54)

    # --- 分支 55-75 ---
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 20): triggered.add(55)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 12.9): triggered.add(56)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(20 - z) < 5): triggered.add(57)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - 80) > 10 and abs(x - z) < 5): triggered.add(58)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x + z) < 5): triggered.add(59)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 7): triggered.add(60)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 15 and abs(10 - z) < 10): triggered.add(61)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - 8) < 5): triggered.add(62)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - 2) < 5): triggered.add(63)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x * 4 - z) < 5): triggered.add(64)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x * z) < 5): triggered.add(65)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x + z) < 5): triggered.add(66)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x * y - z) < 5): triggered.add(67)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 2): triggered.add(68)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - y) < 5): triggered.add(69)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 15): triggered.add(70)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x * 2 - z) < 5): triggered.add(71)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(y - z) < 5): triggered.add(72)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x * z - z) < 5): triggered.add(73)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z * 2) < 5): triggered.add(74)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x * x - z) < 5): triggered.add(75)

    # --- 分支 76-85 ---
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x * x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(76)
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x * y > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(77)
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x * z > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(78)
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x > 90 or x < 5) and (y * y > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(79)
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x > 90 or x < 5) and (y * z > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(80)
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x > 90 or x < 5) and (y * x > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(81)
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x * 10 > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(82)
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x * 15 > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(83)
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x > 50 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(84)
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x > 90 or x < 5) and (y > 80 or y < 3) and (z * 40 > 75 or z < 2)): triggered.add(85)

    # --- 分支 86-100 ---
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * x) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): triggered.add(86)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * 60) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): triggered.add(87)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): triggered.add(88)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((70 * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): triggered.add(89)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) + (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): triggered.add(90)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 70 and x ** 2 + y ** 2 > z ** 2): triggered.add(91)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 80 and x ** 2 + y ** 2 > z ** 2): triggered.add(92)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 1.5 + y ** 2 > z ** 2): triggered.add(93)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2.2): triggered.add(94)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2.5): triggered.add(95)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 1.5 > z ** 2): triggered.add(96)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 80 and x ** 2 + y ** 2 > z ** 2): triggered.add(97)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 65 and x ** 2 + y ** 2 > z ** 2): triggered.add(98)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 1.2 + y ** 2 > z ** 2): triggered.add(99)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 1.2 > z ** 2): triggered.add(100)

    # --- 分支 101-105 ---
    if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.7 > (x + y) / 2 and x * y * z > 1000): triggered.add(101)
    if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.5 > (x + y) / 6 and x * y * z > 1000): triggered.add(102)
    if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.8 > (x + y) / 2 and x * y * z > 1000): triggered.add(103)
    if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.5 > (x + x) / 2 and x * y * z > 1000): triggered.add(104)
    if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.5 > (y + y) / 2 and x * y * z > 1000): triggered.add(105)

    return triggered


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0


targetPaths = [
    {15, 20, 29, 31, 32, 38, 41, 42, 45, 48, 49, 50, 57, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 79,
     80, 81, 86, 87, 90},  # A1
    {6, 10, 15, 20, 23, 29, 31, 32, 38, 41, 42, 45, 48, 49, 50, 57, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73,
     74, 75, 79, 80, 81},  # A2
    {1, 3, 4, 5, 6, 8, 9, 10, 13, 14, 57, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 76, 77, 78, 82,
     83, 84, 94, 95, 100},  # A3
    {16, 17, 18, 19, 21, 22, 23, 29, 31, 32, 38, 41, 45, 48, 49, 50, 57, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72,
     73, 74, 75, 105},  # A4
    {6, 7, 10, 11, 15, 57, 58, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 79, 80, 81, 92, 93, 94, 95,
     96, 97, 99, 100},  # A5
    {1, 2, 3, 4, 5, 8, 9, 10, 13, 14, 16, 17, 18, 19, 20, 21, 23, 76, 77, 78, 82, 83, 84, 87, 91, 92, 93, 94, 95, 96,
     97, 98, 99, 100},  # A6
    {11, 15, 20, 57, 59, 61, 62, 63, 64, 65, 66, 67, 69, 71, 72, 73, 74, 75, 79, 80, 81, 88, 89, 91, 92, 93, 94, 95, 97,
     98, 99, 100},  # A7
    {1, 3, 4, 5, 6, 8, 9, 10, 12, 13, 14, 18, 22, 57, 58, 59, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 86, 87,
     91, 92, 97, 98},  # A8
    {16, 18, 19, 21, 22, 23, 24, 25, 29, 31, 32, 35, 38, 48, 50, 57, 59, 61, 62, 63, 64, 65, 66, 67, 69, 71, 72, 73, 74,
     75, 105},  # A9
    {15, 20, 29, 31, 32, 38, 41, 42, 45, 48, 49, 50, 59, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 87, 90,
     102, 105},  # A10
    {1, 2, 3, 4, 5, 8, 9, 10, 12, 13, 14, 19, 23, 24, 25, 26, 27, 29, 30, 31, 32, 33, 35, 37, 46, 47, 53, 101, 102,
     103},  # A11
    {3, 4, 5, 6, 8, 10, 12, 13, 14, 18, 22, 57, 59, 62, 64, 65, 66, 67, 68, 69, 72, 73, 74, 85, 86, 87, 91, 92, 97, 98},
    # A12
    {15, 20, 23, 29, 31, 32, 37, 43, 44, 46, 47, 53, 54, 57, 59, 61, 62, 63, 64, 65, 66, 67, 69, 71, 72, 73, 74, 75},
    # A13
    {6, 10, 16, 18, 19, 21, 22, 23, 25, 29, 31, 32, 35, 38, 45, 48, 49, 50, 55, 56, 60, 70, 79, 80, 81, 105},  # A14
    {3, 4, 5, 6, 8, 9, 10, 12, 13, 14, 17, 19, 20, 21, 27, 30, 33, 55, 57, 61, 63, 64, 70, 88, 89, 90, 104},  # A15
    {2, 3, 4, 5, 8, 9, 10, 14, 19, 23, 24, 26, 27, 28, 30, 33, 62, 63, 67, 101, 102, 103, 104},  # A16
    {1, 2, 3, 4, 5, 8, 9, 10, 12, 13, 14, 19, 28, 34, 36, 37, 46, 47, 53, 101, 102, 103},  # A17
    {15, 20, 29, 31, 32, 37, 39, 40, 51, 52, 55, 56, 70, 79, 80, 81, 86, 87, 90}  # A18
]


class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 6000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(targetPaths)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }


def compute_robustness(state, path):
    base = execute_Tr(*state)
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                        np.clip(state[0] + dx, STATE_MIN_X, STATE_MAX_X),
                        np.clip(state[1] + dy, STATE_MIN_Y, STATE_MAX_Y),
                        np.clip(state[2] + dz, STATE_MIN_Z, STATE_MAX_Z)
                    ])
                n_trig = execute_Tr(*neighbor)
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = targetPaths[target_path_idx]
    samples = []
    attempts = 0

    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(STATE_MIN_X, STATE_MAX_X),
            random.randint(STATE_MIN_Y, STATE_MAX_Y),
            random.randint(STATE_MIN_Z, STATE_MAX_Z)
        ])
        triggered = execute_Tr(*state)

        if not triggered:
            continue

        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)

        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }

        samples.append(sample_data)

    return samples


def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample[
                'robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)

        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]

    return selected_samples


def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = targetPaths[target_path_idx]
        samples = []
        attempts = 0

        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(STATE_MIN_X, STATE_MAX_X),
                random.randint(STATE_MIN_Y, STATE_MAX_Y),
                random.randint(STATE_MIN_Z, STATE_MAX_Z)
            ])
            triggered = execute_Tr(*state)

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)

            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)

        random.shuffle(samples)
        return samples[:config.top_k_samples]

    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")

        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)


def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}

    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates

    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}

        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config,
                                                         shared_candidates[path_idx])

            strategy_results[path_idx] = samples

        results[strategy_name] = strategy_results

    return results


def analyze_fitness_values(results, config):
    analysis_results = {}

    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []

        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])

        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }

        all_scores = []

        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])

        scores_array = np.array(all_scores)

        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)

        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })

        analysis_results[strategy_name] = analysis

    return analysis_results


def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []

    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)

    df = pd.DataFrame(df_data)
    return df, analysis_results


def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []

    print(f"Starting {num_runs} experiments...")

    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")

        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)

        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)

        print(f"Experiment {run_idx + 1} completed.")

    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df


def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')

    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width

    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []

    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]

        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)

    stats_df = pd.DataFrame(stats_data)

    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Results saved to: {output_path}")


def main():
    results_df = run_multiple_experiments(num_runs=20)

    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")

    save_results_to_excel(results_df, output_path)

    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)

    return results_df, output_path


if __name__ == "__main__":
    results, output_file = main()