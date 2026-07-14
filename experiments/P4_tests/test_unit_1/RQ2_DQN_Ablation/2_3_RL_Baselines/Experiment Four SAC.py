import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
from torch.distributions import Normal
import numpy as np
import random
import time
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os
import warnings
warnings.filterwarnings('ignore')

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# ===  ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'ACTION_DIM': 3,
    # : [co2, moisture, temp]
    'MIN_VALUES': np.array([2, 10, 1], dtype=np.float32),
    'MAX_VALUES': np.array([100, 105, 110], dtype=np.float32),
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 3,
    'REPLAY_BATCH_SIZE': 64,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'ACTOR_LR': 3e-4,
    'CRITIC_LR': 3e-4,
    'ALPHA_LR': 3e-4,
    'GAMMA': 0.99,
    'TAU': 0.005,
    'NUM_RUNS': 20,
    'TOP_K_SAMPLES': 20,
    'TARGET_PATHS': [
        {15, 20, 29, 31, 32, 38, 41, 42, 45, 48, 49, 50, 57, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 79,
     80, 81, 86, 87, 90},  # A1
    {6, 10, 15, 20, 23, 29, 31, 32, 38, 41, 42, 45, 48, 49, 50, 57, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73,
     74, 75, 79, 80, 81},  # A2
    {1, 3, 4, 5, 6, 8, 9, 10, 13, 14, 57, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 76, 77, 78, 82,
     83, 84, 94, 95, 100},  # A3
    {16, 17, 18, 19, 21, 22, 23, 29, 31, 32, 38, 41, 45, 48, 49, 50, 57, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72,
     73, 74, 75, 105},  # A4
    {6, 7, 10, 11, 15, 57, 58, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 79, 80, 81, 92, 93, 94, 95,
     96, 97, 99, 100},  # A5
    {1, 2, 3, 4, 5, 8, 9, 10, 13, 14, 16, 17, 18, 19, 20, 21, 23, 76, 77, 78, 82, 83, 84, 87, 91, 92, 93, 94, 95, 96,
     97, 98, 99, 100},  # A6
    {11, 15, 20, 57, 59, 61, 62, 63, 64, 65, 66, 67, 69, 71, 72, 73, 74, 75, 79, 80, 81, 88, 89, 91, 92, 93, 94, 95, 97,
     98, 99, 100},  # A7
    {1, 3, 4, 5, 6, 8, 9, 10, 12, 13, 14, 18, 22, 57, 58, 59, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 86, 87,
     91, 92, 97, 98},  # A8
    {16, 18, 19, 21, 22, 23, 24, 25, 29, 31, 32, 35, 38, 48, 50, 57, 59, 61, 62, 63, 64, 65, 66, 67, 69, 71, 72, 73, 74,
     75, 105},  # A9
    {15, 20, 29, 31, 32, 38, 41, 42, 45, 48, 49, 50, 59, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 87, 90,
     102, 105},  # A10
    {1, 2, 3, 4, 5, 8, 9, 10, 12, 13, 14, 19, 23, 24, 25, 26, 27, 29, 30, 31, 32, 33, 35, 37, 46, 47, 53, 101, 102,
     103},  # A11
    {3, 4, 5, 6, 8, 10, 12, 13, 14, 18, 22, 57, 59, 62, 64, 65, 66, 67, 68, 69, 72, 73, 74, 85, 86, 87, 91, 92, 97, 98},
    # A12
    {15, 20, 23, 29, 31, 32, 37, 43, 44, 46, 47, 53, 54, 57, 59, 61, 62, 63, 64, 65, 66, 67, 69, 71, 72, 73, 74, 75},
    # A13
    {6, 10, 16, 18, 19, 21, 22, 23, 25, 29, 31, 32, 35, 38, 45, 48, 49, 50, 55, 56, 60, 70, 79, 80, 81, 105},  # A14
    {3, 4, 5, 6, 8, 9, 10, 12, 13, 14, 17, 19, 20, 21, 27, 30, 33, 55, 57, 61, 63, 64, 70, 88, 89, 90, 104},  # A15
    {2, 3, 4, 5, 8, 9, 10, 14, 19, 23, 24, 26, 27, 28, 30, 33, 62, 63, 67, 101, 102, 103, 104},  # A16
    {1, 2, 3, 4, 5, 8, 9, 10, 12, 13, 14, 19, 28, 34, 36, 37, 46, 47, 53, 101, 102, 103},  # A17
    {15, 20, 29, 31, 32, 37, 39, 40, 51, 52, 55, 56, 70, 79, 80, 81, 86, 87, 90}  # A18
    ],
}


# ===  ===
def clip_state(state):
    """"""
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return np.clip(state, min_vals, max_vals)


def normalize_state(state):
    """[-1, 1]"""
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return 2 * (state - min_vals) / (max_vals - min_vals) - 1


def denormalize_state(normalized_state):
    """"""
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return (normalized_state + 1) * (max_vals - min_vals) / 2 + min_vals


def coverage_similarity(triggered, target_path):
    if triggered is None:
        return 0.0
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0

    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)


def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']

    if triggered is not None and target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']

    if triggered is not None and len(triggered) > 0:
        reward += config['TRIGGER_BONUS']

    return reward


def execute_Tr(x, y, z):
    """执行测试覆盖，返回触发的分支集合"""
    # 使用浮点数进行计算，避免整数溢出
    x, y, z = float(x), float(y), float(z)
    
    triggered = set()
    
    # 使用try-except处理可能的数学错误
    try:
        # --- 分支 1-14 ---
        if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 1) > 120): triggered.add(1)
        if ((y * z) / (x + 1) > 100) != ((y * y) / (x + 1) > 100): triggered.add(2)
        if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 8) > 100): triggered.add(3)
        if ((y * z) / (x + 1) > 100) != ((y + z) / (x + 1) > 100): triggered.add(4)
        if ((y * z) / (x + 1) > 100) != ((y * z) / (x * 3 + 1) > 100): triggered.add(5)
        if ((y * z) / (x + 1) > 100) != ((z * z) / (x + 1) > 100): triggered.add(6)
        if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 1) > 80): triggered.add(7)
        if ((y * z) / (x + 1) > 100) != ((y * z) / (x + y) > 100): triggered.add(8)
        if ((y * z) / (x + 1) > 100) != ((y * 10) / (x + 1) > 100): triggered.add(9)
        if ((y * z) / (x + 1) > 100) != ((x * z) / (x + 1) > 100): triggered.add(10)
        if ((y * z) / (x + 1) > 100) != ((y * z * 2) / (x + 1) > 100): triggered.add(11)
        if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 5) > 100): triggered.add(12)
        if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 1) > 200): triggered.add(13)
        if ((y * z) / (x + 1) > 100) != ((5 * z) / (x + 1) > 100): triggered.add(14)

        # --- 分支 15-22 ---
        if ((z - x) < 0.3 * y) != ((z * 1.2 - x) < 0.3 * y): triggered.add(15)
        if ((z - x) < 0.3 * y) != ((z - x * 1.1) < 0.3 * y): triggered.add(16)
        if ((z - x) < 0.3 * y) != ((z - x) < 0.5 * y): triggered.add(17)
        if ((z - x) < 0.3 * y) != ((z - x) < 0.3 * z): triggered.add(18)
        if ((z - x) < 0.3 * y) != ((z - x) < 0.3 * y * y): triggered.add(19)
        if ((z - x) < 0.3 * y) != ((z - y) < 0.3 * y): triggered.add(20)
        if ((z - x) < 0.3 * y) != ((z - x) < 0.8 * y): triggered.add(21)
        if ((z - x) < 0.3 * y) != ((z - x) < 0.3 * x): triggered.add(22)

        # --- 分支 23-36 (处理可能的溢出) ---
        try:
            if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 3): triggered.add(23)
        except: pass
        try:
            if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 * y ** 3) < z ** 4): triggered.add(24)
        except: pass
        if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2 + y ** 3) < z ** 2): triggered.add(25)
        if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 2) < z ** 2): triggered.add(26)
        if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * 3) < (z ** 2)): triggered.add(27)
        if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * x ** 3) < z ** 2): triggered.add(28)
        if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 1 + y ** 3) < z ** 2): triggered.add(29)
        if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 1) < z ** 2): triggered.add(30)
        if ((x ** 3 + y ** 3) < z ** 2) != ((x * 3 + y ** 3) < z ** 2): triggered.add(31)
        if ((x ** 3 + y ** 3) < z ** 2) != ((x * 3 + y ** 3) < z ** 2): triggered.add(32)
        if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * 3) < z ** 2): triggered.add(33)
        try:
            # 防止 x**x 溢出，使用对数判断
            if x > 0 and x < 10:  # 只在安全范围内计算
                if ((x ** 3 + y ** 3) < z ** 2) != ((x ** x + y ** 3) < z ** 2): triggered.add(34)
        except: pass
        if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2 + y ** 3) < z ** 2): triggered.add(35)
        if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 1.5): triggered.add(36)

        # --- 分支 37-54 ---
        if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 or (y / (z + 0.001) < 0.2)): triggered.add(37)
        if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y % (z + 0.001) < 0.2)): triggered.add(38)
        if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y / (z - 0.001) < 0.4)): triggered.add(39)
        if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.3)): triggered.add(40)
        if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.1)): triggered.add(41)
        if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 8 and (y / (z + 0.001) < 0.2)): triggered.add(42)
        if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 2 and (y / (z + 0.001) < 0.2)): triggered.add(43)
        if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 1 and (y / (z + 0.001) < 0.2)): triggered.add(44)
        if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 10 and (y / (z + 0.001) < 0.2)): triggered.add(45)
        if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x * y / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.2)): triggered.add(46)
        if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x * x / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.2)): triggered.add(47)
        if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y * x / (z + 0.001) < 0.2)): triggered.add(48)
        if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y * y / (z + 0.001) < 0.2)): triggered.add(49)
        if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y * z / (z + 0.001) < 0.2)): triggered.add(50)
        if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.5)): triggered.add(51)
        if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y - (z + 0.001) < 0.2)): triggered.add(52)
        if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x + (y + 0.001)) > 5 and (y / (z + 0.001) < 0.2)): triggered.add(53)
        if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x - (y + 0.001)) > 5 and (y / (z + 0.001) < 0.2)): triggered.add(54)

        # --- 分支 55-75 ---
        if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 20): triggered.add(55)
        if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 12.9): triggered.add(56)
        if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(20 - z) < 5): triggered.add(57)
        if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - 80) > 10 and abs(x - z) < 5): triggered.add(58)
        if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x + z) < 5): triggered.add(59)
        if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 7): triggered.add(60)
        if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 15 and abs(10 - z) < 10): triggered.add(61)
        if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - 8) < 5): triggered.add(62)
        if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - 2) < 5): triggered.add(63)
        if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x * 4 - z) < 5): triggered.add(64)
        if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x * z) < 5): triggered.add(65)
        if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x + z) < 5): triggered.add(66)
        if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x * y - z) < 5): triggered.add(67)
        if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 2): triggered.add(68)
        if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - y) < 5): triggered.add(69)
        if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 15): triggered.add(70)
        if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x * 2 - z) < 5): triggered.add(71)
        if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(y - z) < 5): triggered.add(72)
        if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x * z - z) < 5): triggered.add(73)
        if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z * 2) < 5): triggered.add(74)
        if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x * x - z) < 5): triggered.add(75)

        # --- 分支 76-85 ---
        if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x * x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(76)
        if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x * y > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(77)
        if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x * z > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(78)
        if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x > 90 or x < 5) and (y * y > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(79)
        if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x > 90 or x < 5) and (y * z > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(80)
        if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x > 90 or x < 5) and (y * x > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(81)
        if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x * 10 > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(82)
        if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x * 15 > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(83)
        if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x > 50 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(84)
        if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x > 90 or x < 5) and (y > 80 or y < 3) and (z * 40 > 75 or z < 2)): triggered.add(85)

        # --- 分支 86-100 ---
        if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * x) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): triggered.add(86)
        if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * 60) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): triggered.add(87)
        if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): triggered.add(88)
        if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((70 * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): triggered.add(89)
        if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) + (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): triggered.add(90)
        if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 70 and x ** 2 + y ** 2 > z ** 2): triggered.add(91)
        if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 80 and x ** 2 + y ** 2 > z ** 2): triggered.add(92)
        if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 1.5 + y ** 2 > z ** 2): triggered.add(93)
        if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2.2): triggered.add(94)
        if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2.5): triggered.add(95)
        if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 1.5 > z ** 2): triggered.add(96)
        if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 80 and x ** 2 + y ** 2 > z ** 2): triggered.add(97)
        if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 65 and x ** 2 + y ** 2 > z ** 2): triggered.add(98)
        if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 1.2 + y ** 2 > z ** 2): triggered.add(99)
        if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 1.2 > z ** 2): triggered.add(100)

        # --- 分支 101-105 ---
        if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.7 > (x + y) / 2 and x * y * z > 1000): triggered.add(101)
        if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.5 > (x + y) / 6 and x * y * z > 1000): triggered.add(102)
        if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.8 > (x + y) / 2 and x * y * z > 1000): triggered.add(103)
        if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.5 > (x + x) / 2 and x * y * z > 1000): triggered.add(104)
        if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.5 > (y + y) / 2 and x * y * z > 1000): triggered.add(105)
        
    except Exception as e:
        # 如果发生任何计算错误，返回已触发的分支
        pass
    
    return triggered


# === SAC Actor ===
class GaussianPolicy(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dim=256):
        super(GaussianPolicy, self).__init__()

        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)

        self.mean = nn.Linear(hidden_dim, action_dim)
        self.log_std = nn.Linear(hidden_dim, action_dim)

        self.action_scale = 8.0
        self.action_bias = 0.0

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))

        mean = self.mean(x)
        log_std = self.log_std(x)
        log_std = torch.clamp(log_std, min=-20, max=2)

        return mean, log_std

    def sample(self, state):
        mean, log_std = self.forward(state)
        std = log_std.exp()
        normal = Normal(mean, std)

        x_t = normal.rsample()
        action = torch.tanh(x_t) * self.action_scale + self.action_bias

        log_prob = normal.log_prob(x_t)
        log_prob -= torch.log(self.action_scale * (1 - torch.tanh(x_t).pow(2)) + 1e-6)
        log_prob = log_prob.sum(1, keepdim=True)

        mean = torch.tanh(mean) * self.action_scale + self.action_bias

        return action, log_prob, mean


# === SAC Critic ===
class QNetwork(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dim=256):
        super(QNetwork, self).__init__()

        self.fc1 = nn.Linear(state_dim + action_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, 1)

        self.fc4 = nn.Linear(state_dim + action_dim, hidden_dim)
        self.fc5 = nn.Linear(hidden_dim, hidden_dim)
        self.fc6 = nn.Linear(hidden_dim, 1)

    def forward(self, state, action):
        xu = torch.cat([state, action], 1)

        x1 = F.relu(self.fc1(xu))
        x1 = F.relu(self.fc2(x1))
        q1 = self.fc3(x1)

        x2 = F.relu(self.fc4(xu))
        x2 = F.relu(self.fc5(x2))
        q2 = self.fc6(x2)

        return q1, q2


# ===  ===
class EnhancedReplayBuffer:
    def __init__(self, capacity=50000):
        self.buffer = deque(maxlen=capacity)
        self.experience_info = deque(maxlen=capacity)

    def push(self, state, action, reward, next_state, done, path_idx, similarity):
        self.buffer.append((state, action, reward, next_state, done))
        self.experience_info.append({'path_idx': path_idx, 'similarity': similarity})

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return None

        batch = random.sample(self.buffer, batch_size)
        state, action, reward, next_state, done = map(np.stack, zip(*batch))

        return (
            torch.FloatTensor(state).to(device),
            torch.FloatTensor(action).to(device),
            torch.FloatTensor(reward).unsqueeze(1).to(device),
            torch.FloatTensor(next_state).to(device),
            torch.FloatTensor(done).unsqueeze(1).to(device)
        )

    def get_top_k_per_path(self, num_paths, k=20):
        path_samples = {i: [] for i in range(num_paths)}

        for idx, info in enumerate(self.experience_info):
            path_idx = info['path_idx']
            similarity = info['similarity']
            path_samples[path_idx].append((idx, similarity, self.buffer[idx]))

        top_k_results = {}
        for path_idx in range(num_paths):
            samples = path_samples[path_idx]
            if len(samples) == 0:
                top_k_results[path_idx] = []
                continue

            samples.sort(key=lambda x: x[1], reverse=True)
            top_k = samples[:k]

            top_k_results[path_idx] = []
            for sample in top_k:
                normalized_state = sample[2][0]
                original_state = denormalize_state(normalized_state)
                original_state_int = np.round(original_state).astype(int)

                # 执行测试覆盖
                co2, moisture, temp = original_state_int
                triggered = execute_Tr(co2, moisture, temp)

                top_k_results[path_idx].append({
                    'state': original_state_int,
                    'similarity': sample[1],
                    'triggered': triggered
                })

        return top_k_results

    def __len__(self):
        return len(self.buffer)


# === SAC ===
class SACAgent:
    def __init__(self, state_dim=3, action_dim=3):
        self.state_dim = state_dim
        self.action_dim = action_dim

        # 初始化网络
        self.policy = GaussianPolicy(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.policy_optimizer = optim.Adam(self.policy.parameters(), lr=EXPERIMENT_CONFIG['ACTOR_LR'])

        self.critic = QNetwork(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_target = QNetwork(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_target.load_state_dict(self.critic.state_dict())
        self.critic_optimizer = optim.Adam(self.critic.parameters(), lr=EXPERIMENT_CONFIG['CRITIC_LR'])

        self.target_entropy = -action_dim
        self.log_alpha = torch.zeros(1, requires_grad=True, device=device)
        self.alpha_optimizer = optim.Adam([self.log_alpha], lr=EXPERIMENT_CONFIG['ALPHA_LR'])

        self.replay_buffer = EnhancedReplayBuffer()
        self.replay_train_count = 0

    def get_action(self, state, deterministic=False):
        normalized_state = normalize_state(state)
        state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)

        with torch.no_grad():
            if deterministic:
                _, _, action = self.policy.sample(state_tensor)
            else:
                action, _, _ = self.policy.sample(state_tensor)

        action = action.cpu().numpy()[0]
        return action

    def store_experience(self, state, action, reward, next_state, done, path_idx, similarity):
        normalized_state = normalize_state(state)
        normalized_next_state = normalize_state(next_state)

        self.replay_buffer.push(
            normalized_state, action, reward,
            normalized_next_state, done, path_idx, similarity
        )

    def replay_train(self):
        batch = self.replay_buffer.sample(EXPERIMENT_CONFIG['REPLAY_BATCH_SIZE'])

        if batch is None:
            return

        state, action, reward, next_state, done = batch

        with torch.no_grad():
            next_action, next_log_prob, _ = self.policy.sample(next_state)

            q1_next, q2_next = self.critic_target(next_state, next_action)
            q_next = torch.min(q1_next, q2_next)
            target_q = reward + (1 - done) * EXPERIMENT_CONFIG['GAMMA'] * (
                    q_next - self.log_alpha.exp() * next_log_prob)

        q1, q2 = self.critic(state, action)
        critic_loss = F.mse_loss(q1, target_q) + F.mse_loss(q2, target_q)

        self.critic_optimizer.zero_grad()
        critic_loss.backward()
        self.critic_optimizer.step()

        new_action, log_prob, _ = self.policy.sample(state)
        q1_new, q2_new = self.critic(state, new_action)
        q_new = torch.min(q1_new, q2_new)

        policy_loss = (self.log_alpha.exp() * log_prob - q_new).mean()

        self.policy_optimizer.zero_grad()
        policy_loss.backward()
        self.policy_optimizer.step()

        alpha_loss = -(self.log_alpha * (log_prob + self.target_entropy).detach()).mean()

        self.alpha_optimizer.zero_grad()
        alpha_loss.backward()
        self.alpha_optimizer.step()

        for param, target_param in zip(self.critic.parameters(), self.critic_target.parameters()):
            target_param.data.copy_(
                EXPERIMENT_CONFIG['TAU'] * param.data + (1 - EXPERIMENT_CONFIG['TAU']) * target_param.data)

        self.replay_train_count += 1

        if self.replay_train_count % 2 == 0:
            alpha_value = self.log_alpha.exp().item()
            print(f"  -> 训练更新 (Run {self.replay_train_count}), Alpha={alpha_value:.4f}")


# === Metric ===
def calculate_run_performance(run_idx, sac_results, training_time, total_steps, update_count, agent):
    """计算单个run的性能指标"""
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    # 收集所有相似度
    all_similarities = []

    # 收集所有奖励
    total_samples = 0
    all_rewards = []
    total_reward = 0

    for path_idx in range(num_paths):
        samples = sac_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']

            total_reward += reward
            all_rewards.append(reward)
            all_similarities.append(similarity)
            total_samples += 1

    # 1. 总奖励
    total_reward = total_reward

    # 2. 平均奖励
    if total_samples > 0:
        average_reward = total_reward / total_samples
    else:
        average_reward = 0

    # 5. 收敛性 (平均相似度)
    if all_similarities:
        convergence = np.mean(all_similarities)
    else:
        convergence = 0

    # 12. 环境适应性 (相似度的标准差倒数)
    if len(all_similarities) > 1:
        environment_adaptability = 1 / (np.std(all_similarities) + 1e-8)
    else:
        environment_adaptability = 0

    # 13. 泛化能力 (平均相似度)
    generalization_ability = convergence

    # 15. 计算效率 (总步数 / 训练时间)
    if training_time > 0:
        computational_efficiency = total_steps / training_time
    else:
        computational_efficiency = 0

    # 16. 策略更新频率
    if training_time > 0:
        policy_update_frequency = update_count / training_time
    else:
        policy_update_frequency = 0

    # 相似度统计
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0

    return {
        'Run': run_idx + 1,

        # 性能指标
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),

        # 相似度统计
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }


# === Excel导出 ===
def export_to_excel(all_sac_results, all_performance_data, target_paths, output_path="SAC_20_run.xlsx"):
    """导出 20 次 run 的 SAC 测试结果到 Excel"""
    print("\n正在导出数据到 Excel...")

    # 初始化收集所有 run 的数据
    all_sac_summary_data = []
    all_sac_detailed_data = []

    # 遍历每次 run 的结果
    for run_idx, (sac_results, performance_data) in enumerate(zip(all_sac_results, all_performance_data)):
        # ===== Sheet1: SACPath 统计 =====
        sac_summary_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = sac_results[path_idx]

            if len(samples) == 0:
                sac_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue

            # 正常结果计算
            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = 'Yes' if perfect_count > 0 else 'No'

            sac_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })

        all_sac_summary_data.extend(sac_summary_data)

        # ===== Sheet2: SACDetailed Sample Data =====
        sac_detailed_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = sac_results[path_idx]

            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']

                sac_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Sample ID': sample_idx + 1,
                    'CO2': int(state[0]),
                    'Moisture': int(state[1]),
                    'Temperature': int(state[2]),
                    'Similarity': round(similarity, 4),
                    'Perfect Coverage': 'Yes' if abs(similarity - 1.0) < 0.001 else 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path))),
                    'Triggered Rules': ', '.join(map(str, sorted(triggered))) if triggered else '',
                    'Intersection Count': len(target_path.intersection(triggered)) if triggered else 0,
                    'Target Rule Count': len(target_path)
                })

        all_sac_detailed_data.extend(sac_detailed_data)

    # 转换为 DataFrame
    sac_summary_df = pd.DataFrame(all_sac_summary_data)
    sac_detailed_df = pd.DataFrame(all_sac_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet1: SACPath 统计
        sac_summary_df.to_excel(writer, sheet_name='SACPath', index=False)

        # Sheet2: SACDetailed Sample Data
        sac_detailed_df.to_excel(writer, sheet_name='SACDetailed Sample Data', index=False)

        # Sheet3: Metric - 性能指标汇总
        selected_columns = [
            'Run',
            'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df_selected = performance_df[selected_columns]
        performance_df_selected.to_excel(writer, sheet_name='Metric', index=False)

        # 获取 workbook 对象进行格式设置
        workbook = writer.book

        # 设置表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        perfect_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # 浅绿色高亮

        # === 设置 Sheet1 样式 ===
        ws1 = writer.sheets['SACPath']
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 高亮完全覆盖的行
        for row_idx in range(2, ws1.max_row + 1):
            if ws1.cell(row_idx, 9).value == 'Yes':  # 第9列是 Perfect Coverage
                for col_idx in range(1, ws1.max_column + 1):
                    ws1.cell(row_idx, col_idx).fill = perfect_fill

        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 18
        ws1.column_dimensions['D'].width = 15
        ws1.column_dimensions['E'].width = 18
        ws1.column_dimensions['F'].width = 15
        ws1.column_dimensions['G'].width = 15
        ws1.column_dimensions['H'].width = 15
        ws1.column_dimensions['I'].width = 18
        ws1.column_dimensions['J'].width = 50

        # === 设置 Sheet2 样式 ===
        ws2 = writer.sheets['SACDetailed Sample Data']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 18
        ws2.column_dimensions['I'].width = 40
        ws2.column_dimensions['J'].width = 40
        ws2.column_dimensions['K'].width = 18
        ws2.column_dimensions['L'].width = 18

        # === 设置 Sheet3 样式 ===
        ws3 = writer.sheets['Metric']
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 调整列宽
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        for col in columns:
            ws3.column_dimensions[col].width = 20

    print(f"文件已成功保存到: {output_path}")
    print(f"  - Sheet1: SACPath 统计共计 {len(all_sac_summary_data)} 条记录")
    print(f"  - Sheet2: SACDetailed Sample Data 共计 {len(all_sac_detailed_data)} 条记录")
    print(f"  - Sheet3: Metric 共计 {len(all_performance_data)} 条记录")


# === 训练工作流 ===
def train_sac_workflow():
    print("=" * 80)
    print("开始SAC训练")
    print("Similarity: 触发分支数 / target paths目标分支数")
    print("=" * 80)

    agent = SACAgent()
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n初始化: 为每个Path生成 {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']} 个样本")
    path_samples = {}
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']

    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            # 生成随机初始状态
            state = np.random.uniform(min_vals, max_vals).astype(np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  Path  {path_idx + 1}/{num_paths}:  生成 {len(samples)} 个样本")

    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size

    print(f"\n开始训练: 每批次{batch_size}个样本, 每个样本{EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}步")
    print(f"总批次数: {num_batches} 批次/Path  x {num_paths} Path  = {num_batches * num_paths} 批次")
    print("-" * 80)

    for batch_idx in range(num_batches):
        print(f"\n批次 {batch_idx + 1}/{num_batches}")

        for path_idx in range(num_paths):
            target_path = target_paths[path_idx]
            batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]

            batch_rewards = []
            batch_similarities = []

            for sample_idx, initial_state in enumerate(batch_samples):
                state = initial_state.copy()
                episode_reward = 0
                final_similarity = 0

                for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                    action = agent.get_action(state, deterministic=False)

                    next_state = state + action
                    next_state = clip_state(next_state)

                    # 计算奖励和相似度
                    co2, moisture, temp = next_state
                    triggered = execute_Tr(co2, moisture, temp)
                    reward = unified_reward_function(triggered, target_path)
                    similarity = coverage_similarity(triggered, target_path)

                    done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)

                    agent.store_experience(
                        state, action, reward, next_state, done,
                        path_idx, similarity
                    )

                    state = next_state
                    episode_reward += reward
                    final_similarity = similarity
                    total_steps += 1

                batch_rewards.append(episode_reward)
                batch_similarities.append(final_similarity)

            avg_reward = np.mean(batch_rewards)
            avg_similarity = np.mean(batch_similarities)
            print(f"  Path {path_idx + 1}: 平均奖励={avg_reward:.2f}, 平均相似度={avg_similarity:.4f}")

        print(f"\n  执行训练更新...")
        agent.replay_train()
        print(f"  缓冲区大小: {len(agent.replay_buffer)}")

    training_time = time.time() - start_time

    print("\n" + "=" * 80)
    print(f"SAC训练完成! 总耗时: {training_time:.2f} 秒, 总步数: {total_steps}")
    print(f"缓冲区大小: {len(agent.replay_buffer)}")
    print(f"训练更新次数: {agent.replay_train_count}")
    print("=" * 80)

    print(f"\n获取每个Path相似度最高的{EXPERIMENT_CONFIG['TOP_K_SAMPLES']}个样本...")
    top_k_results = agent.replay_buffer.get_top_k_per_path(num_paths, EXPERIMENT_CONFIG['TOP_K_SAMPLES'])

    return agent, top_k_results, training_time, total_steps, agent.replay_train_count


# === 主函数 ===
def main():
    print("\n" + "=" * 80)
    print("SAC - 20 run实验")
    print("环境: CO2(2-100), moisture(10-105), temperature(1-110)")
    print("评估Metric性能指标")
    print("=" * 80)

    all_sac_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    # 执行20次run
    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'=' * 80}")
        print(f"开始执行第 {run_idx + 1}/{EXPERIMENT_CONFIG['NUM_RUNS']} 次run")
        print(f"{'=' * 80}")

        # 执行SAC训练
        sac_agent, sac_results, training_time, total_steps, update_count = train_sac_workflow()

        # 计算性能指标
        performance_data = calculate_run_performance(
            run_idx, sac_results, training_time, total_steps, update_count, sac_agent
        )

        # 存储结果
        all_sac_results.append(sac_results)
        all_performance_data.append(performance_data)

        print(f"\n第 {run_idx + 1} 次run完成!")
        print(f"  总奖励: {performance_data['Total Reward']}")
        print(f"  平均奖励: {performance_data['Average Reward']}")
        print(f"  平均相似度: {performance_data['Average Similarity']}")

    # 导出所有结果到Excel(包含20次run)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"SAC_20_run_{timestamp}.xlsx"
    export_to_excel(all_sac_results, all_performance_data, target_paths, output_path)

    # 汇总统计
    print("\n" + "=" * 80)
    print("20次run汇总统计")
    print("=" * 80)

    # 提取所有性能指标
    total_rewards = [p['Total Reward'] for p in all_performance_data]
    average_rewards = [p['Average Reward'] for p in all_performance_data]
    convergences = [p['Convergence'] for p in all_performance_data]
    environment_adaptabilities = [p['Environment Adaptability'] for p in all_performance_data]
    generalization_abilities = [p['Generalization Ability'] for p in all_performance_data]
    computational_efficiencies = [p['Computational Efficiency'] for p in all_performance_data]
    policy_update_frequencies = [p['Policy Update Frequency'] for p in all_performance_data]
    avg_similarities = [p['Average Similarity'] for p in all_performance_data]

    print(f"总奖励统计:")
    print(f"  平均值: {np.mean(total_rewards):.2f}")
    print(f"  标准差: {np.std(total_rewards):.2f}")

    print(f"\n平均奖励统计:")
    print(f"  平均值: {np.mean(average_rewards):.4f}")
    print(f"  标准差: {np.std(average_rewards):.4f}")

    print(f"\n收敛性统计:")
    print(f"  平均值: {np.mean(convergences):.4f}")
    print(f"  标准差: {np.std(convergences):.4f}")

    print(f"\n环境适应性统计:")
    print(f"  平均值: {np.mean(environment_adaptabilities):.4f}")
    print(f"  标准差: {np.std(environment_adaptabilities):.4f}")

    print(f"\n泛化能力统计:")
    print(f"  平均值: {np.mean(generalization_abilities):.4f}")
    print(f"  标准差: {np.std(generalization_abilities):.4f}")

    print(f"\n计算效率统计:")
    print(f"  平均值: {np.mean(computational_efficiencies):.2f}")
    print(f"  标准差: {np.std(computational_efficiencies):.2f}")

    print(f"\n策略更新频率统计:")
    print(f"  平均值: {np.mean(policy_update_frequencies):.4f}")
    print(f"  标准差: {np.std(policy_update_frequencies):.4f}")

    print(f"\n平均相似度统计:")
    print(f"  平均值: {np.mean(avg_similarities):.4f}")
    print(f"  标准差: {np.std(avg_similarities):.4f}")

    print("\n" + "=" * 80)
    print(f"全部 {EXPERIMENT_CONFIG['NUM_RUNS']} 次run已完成!")
    print("=" * 80)


if __name__ == "__main__":
    main()