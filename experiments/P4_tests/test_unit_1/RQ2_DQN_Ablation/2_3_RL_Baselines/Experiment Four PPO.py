import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
import numpy as np
import random
import time
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
from datetime import datetime

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# ===  ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'ACTION_DIM': 3,
    'MIN_VALUES': np.array([2, 1, 1]),
    'MAX_VALUES': np.array([100, 105, 110]),
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 3,
    'PPO_BATCH_SIZE': 64,
    'PPO_EPOCHS': 4,
    'CLIP_EPSILON': 0.2,
    'GAMMA': 0.99,
    'GAE_LAMBDA': 0.95,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'ACTOR_LR': 3e-4,
    'CRITIC_LR': 3e-4,
    'NUM_RUNS': 20,
    'TOP_K_SAMPLES': 20,
    'TARGET_PATHS': [
        {15, 20, 29, 31, 32, 38, 41, 42, 45, 48, 49, 50, 57, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 79,
     80, 81, 86, 87, 90},  # A1
    {6, 10, 15, 20, 23, 29, 31, 32, 38, 41, 42, 45, 48, 49, 50, 57, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73,
     74, 75, 79, 80, 81},  # A2
    {1, 3, 4, 5, 6, 8, 9, 10, 13, 14, 57, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 76, 77, 78, 82,
     83, 84, 94, 95, 100},  # A3
    {16, 17, 18, 19, 21, 22, 23, 29, 31, 32, 38, 41, 45, 48, 49, 50, 57, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72,
     73, 74, 75, 105},  # A4
    {6, 7, 10, 11, 15, 57, 58, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 79, 80, 81, 92, 93, 94, 95,
     96, 97, 99, 100},  # A5
    {1, 2, 3, 4, 5, 8, 9, 10, 13, 14, 16, 17, 18, 19, 20, 21, 23, 76, 77, 78, 82, 83, 84, 87, 91, 92, 93, 94, 95, 96,
     97, 98, 99, 100},  # A6
    {11, 15, 20, 57, 59, 61, 62, 63, 64, 65, 66, 67, 69, 71, 72, 73, 74, 75, 79, 80, 81, 88, 89, 91, 92, 93, 94, 95, 97,
     98, 99, 100},  # A7
    {1, 3, 4, 5, 6, 8, 9, 10, 12, 13, 14, 18, 22, 57, 58, 59, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 86, 87,
     91, 92, 97, 98},  # A8
    {16, 18, 19, 21, 22, 23, 24, 25, 29, 31, 32, 35, 38, 48, 50, 57, 59, 61, 62, 63, 64, 65, 66, 67, 69, 71, 72, 73, 74,
     75, 105},  # A9
    {15, 20, 29, 31, 32, 38, 41, 42, 45, 48, 49, 50, 59, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 87, 90,
     102, 105},  # A10
    {1, 2, 3, 4, 5, 8, 9, 10, 12, 13, 14, 19, 23, 24, 25, 26, 27, 29, 30, 31, 32, 33, 35, 37, 46, 47, 53, 101, 102,
     103},  # A11
    {3, 4, 5, 6, 8, 10, 12, 13, 14, 18, 22, 57, 59, 62, 64, 65, 66, 67, 68, 69, 72, 73, 74, 85, 86, 87, 91, 92, 97, 98},
    # A12
    {15, 20, 23, 29, 31, 32, 37, 43, 44, 46, 47, 53, 54, 57, 59, 61, 62, 63, 64, 65, 66, 67, 69, 71, 72, 73, 74, 75},
    # A13
    {6, 10, 16, 18, 19, 21, 22, 23, 25, 29, 31, 32, 35, 38, 45, 48, 49, 50, 55, 56, 60, 70, 79, 80, 81, 105},  # A14
    {3, 4, 5, 6, 8, 9, 10, 12, 13, 14, 17, 19, 20, 21, 27, 30, 33, 55, 57, 61, 63, 64, 70, 88, 89, 90, 104},  # A15
    {2, 3, 4, 5, 8, 9, 10, 14, 19, 23, 24, 26, 27, 28, 30, 33, 62, 63, 67, 101, 102, 103, 104},  # A16
    {1, 2, 3, 4, 5, 8, 9, 10, 12, 13, 14, 19, 28, 34, 36, 37, 46, 47, 53, 101, 102, 103},  # A17
    {15, 20, 29, 31, 32, 37, 39, 40, 51, 52, 55, 56, 70, 79, 80, 81, 86, 87, 90}  # A18
    ],
}

# ===  ===
def execute_Tr(x, y, z):
    # 初始化分支覆盖数组
    b = [0] * 105  # 根据分支数量调整大小

    # --- 分支 1-4 ---
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 1) > 120): b[0] = 1
    if ((y * z) / (x + 1) > 100) != ((y * y) / (x + 1) > 100): b[1] = 2
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 8) > 100): b[2] = 3
    if ((y * z) / (x + 1) > 100) != ((y + z) / (x + 1) > 100): b[3] = 4
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x * 3 + 1) > 100): b[4] = 5
    if ((y * z) / (x + 1) > 100) != ((z * z) / (x + 1) > 100): b[5] = 6
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 1) > 80): b[6] = 7
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + y) > 100): b[7] = 8
    if ((y * z) / (x + 1) > 100) != ((y * 10) / (x + 1) > 100): b[8] = 9
    if ((y * z) / (x + 1) > 100) != ((x * z) / (x + 1) > 100): b[9] = 10
    if ((y * z) / (x + 1) > 100) != ((y * z * 2) / (x + 1) > 100): b[10] = 11
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 5) > 100): b[11] = 12
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 1) > 200): b[12] = 13
    if ((y * z) / (x + 1) > 100) != ((5 * z) / (x + 1) > 100): b[13] = 14

    if ((z - x) < 0.3 * y) != ((z * 1.2 - x) < 0.3 * y): b[14] = 15
    if ((z - x) < 0.3 * y) != ((z - x * 1.1) < 0.3 * y): b[15] = 16
    if ((z - x) < 0.3 * y) != ((z - x) < 0.5 * y): b[16] = 17
    if ((z - x) < 0.3 * y) != ((z - x) < 0.3 * z): b[17] = 18
    if ((z - x) < 0.3 * y) != ((z - x) < 0.3 * y * y): b[18] = 19
    if ((z - x) < 0.3 * y) != ((z - y) < 0.3 * y): b[19] = 20
    if ((z - x) < 0.3 * y) != ((z - x) < 0.8 * y): b[20] = 21
    if ((z - x) < 0.3 * y) != ((z - x) < 0.3 * x): b[21] = 22

    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 3): b[22] = 23
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 * y ** 3) < z ** 4): b[23] = 24
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2 + y ** 3) < z ** 2): b[24] = 25
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 2) < z ** 2): b[25] = 26
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * 3) < (z ** 2)): b[26] = 27
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * x ** 3) < z ** 2): b[27] = 28
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 1 + y ** 3) < z ** 2): b[28] = 29
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 1) < z ** 2): b[29] = 30
    if ((x ** 3 + y ** 3) < z ** 2) != ((x * 3 + y ** 3) < z ** 2): b[30] = 31
    if ((x ** 3 + y ** 3) < z ** 2) != ((x * 3 + y ** 3) < z ** 2): b[31] = 32
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * 3) < z ** 2): b[32] = 33
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** x + y ** 3) < z ** 2): b[33] = 34
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2 + y ** 3) < z ** 2): b[34] = 35
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 1.5): b[35] = 36

    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 or (y / (z + 0.001) < 0.2)): b[
        36] = 37
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y % (z + 0.001) < 0.2)): b[
        37] = 38
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y / (z - 0.001) < 0.4)): b[
        38] = 39
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.3)): b[
        39] = 40
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.1)): b[
        40] = 41
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 8 and (y / (z + 0.001) < 0.2)): b[
        41] = 42
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 2 and (y / (z + 0.001) < 0.2)): b[
        42] = 43
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 1 and (y / (z + 0.001) < 0.2)): b[
        43] = 44
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 10 and (y / (z + 0.001) < 0.2)): b[
        44] = 45
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x * y / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.2)):
        b[45] = 46
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x * x / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.2)):
        b[46] = 47
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y * x / (z + 0.001) < 0.2)):
        b[47] = 48
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y * y / (z + 0.001) < 0.2)):
        b[48] = 49
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y * z / (z + 0.001) < 0.2)):
        b[49] = 50
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.5)): b[
        50] = 51
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y - (z + 0.001) < 0.2)): b[
        51] = 52
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x + (y + 0.001)) > 5 and (y / (z + 0.001) < 0.2)): b[
        52] = 53
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x - (y + 0.001)) > 5 and (y / (z + 0.001) < 0.2)): b[
        53] = 54

    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 20): b[54] = 55
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 12.9): b[55] = 56
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(20 - z) < 5): b[56] = 57
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - 80) > 10 and abs(x - z) < 5): b[57] = 58
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x + z) < 5): b[58] = 59
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 7): b[59] = 60
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 15 and abs(10 - z) < 10): b[60] = 61
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x - 8) < 5): b[61] = 62
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x - 2) < 5): b[62] = 63
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x * 4 - z) < 5): b[63] = 64
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x * z) < 5): b[64] = 65
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x + z) < 5): b[65] = 66
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x * y - z) < 5): b[66] = 67
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 2): b[67] = 68
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x - y) < 5): b[68] = 69
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 15): b[69] = 70
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x * 2 - z) < 5): b[70] = 71
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(y - z) < 5): b[71] = 72
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x * z - z) < 5): b[72] = 73
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z * 2) < 5): b[73] = 74
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x * x - z) < 5): b[74] = 75

    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != (
            (x * x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): b[75] = 76
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != (
            (x * y > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): b[76] = 77
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != (
            (x * z > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): b[77] = 78
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != (
            (x > 90 or x < 5) and (y * y > 80 or y < 3) and (z > 75 or z < 2)): b[78] = 79
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != (
            (x > 90 or x < 5) and (y * z > 80 or y < 3) and (z > 75 or z < 2)): b[79] = 80
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != (
            (x > 90 or x < 5) and (y * x > 80 or y < 3) and (z > 75 or z < 2)): b[80] = 81
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != (
            (x * 10 > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): b[81] = 82
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != (
            (x * 15 > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): b[82] = 83
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != (
            (x > 50 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): b[83] = 84
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != (
            (x > 90 or x < 5) and (y > 80 or y < 3) and (z * 40 > 75 or z < 2)): b[84] = 85

    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * x) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[85] = 86
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * 60) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[86] = 87
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[87] = 88
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((70 * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[88] = 89
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) + (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[89] = 90
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 70 and x ** 2 + y ** 2 > z ** 2):
        b[90] = 91
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 80 and x ** 2 + y ** 2 > z ** 2):
        b[91] = 92
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 1.5 + y ** 2 > z ** 2):
        b[92] = 93
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2.2):
        b[93] = 94
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2.5):
        b[94] = 95
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 1.5 > z ** 2):
        b[95] = 96
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 80 and x ** 2 + y ** 2 > z ** 2):
        b[96] = 97
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 65 and x ** 2 + y ** 2 > z ** 2):
        b[97] = 98
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 1.2 + y ** 2 > z ** 2):
        b[98] = 99
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 1.2 > z ** 2):
        b[99] = 100

    if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.7 > (x + y) / 2 and x * y * z > 1000): b[100] = 101
    if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.5 > (x + y) / 6 and x * y * z > 1000): b[101] = 102
    if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.8 > (x + y) / 2 and x * y * z > 1000): b[102] = 103
    if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.5 > (x + x) / 2 and x * y * z > 1000): b[103] = 104
    if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.5 > (y + y) / 2 and x * y * z > 1000): b[104] = 105

    # 返回触发的分支索引集合
    triggered = set()
    for i, val in enumerate(b):
        if val > 0:
            triggered.add(val)
    return triggered

# === 状态处理辅助函数 ===
def clip_state(state):
    return np.clip(state, EXPERIMENT_CONFIG['MIN_VALUES'], EXPERIMENT_CONFIG['MAX_VALUES'])

def denormalize_state(normalized_state):
    """将归一化状态还原为原始状态"""
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return normalized_state * (max_vals - min_vals) / 2 + (min_vals + max_vals) / 2

def coverage_similarity(triggered, target_path):
    """
    Similarity:  / target paths
    
    """
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0

    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)

def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']

    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']

    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']

    return reward

# === PPO Actor ===
class PPOActor(nn.Module):
    def __init__(self, state_dim=3, action_dim=3, hidden_dim=256):
        super(PPOActor, self).__init__()

        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)

        self.mean_head = nn.Linear(hidden_dim, action_dim)
        self.log_std_head = nn.Linear(hidden_dim, action_dim)

        self.action_scale = 10.0

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))
        x = F.relu(self.fc3(x))

        mean = torch.tanh(self.mean_head(x)) * self.action_scale
        log_std = torch.clamp(self.log_std_head(x), -20, 2)

        return mean, log_std

    def get_action_and_log_prob(self, state):
        mean, log_std = self.forward(state)
        std = torch.exp(log_std)

        dist = torch.distributions.Normal(mean, std)
        action = dist.sample()
        log_prob = dist.log_prob(action).sum(dim=-1)

        return action, log_prob

    def get_log_prob(self, state, action):
        mean, log_std = self.forward(state)
        std = torch.exp(log_std)

        dist = torch.distributions.Normal(mean, std)
        log_prob = dist.log_prob(action).sum(dim=-1)

        return log_prob

# === PPO Critic ===
class PPOCritic(nn.Module):
    def __init__(self, state_dim=3, hidden_dim=256):
        super(PPOCritic, self).__init__()

        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)
        self.value_head = nn.Linear(hidden_dim, 1)

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))
        x = F.relu(self.fc3(x))
        value = self.value_head(x)
        return value.squeeze(-1)

# === PPO ===
class PPOBuffer:
    def __init__(self):
        self.states = []
        self.actions = []
        self.rewards = []
        self.values = []
        self.log_probs = []
        self.dones = []
        self.path_indices = []
        self.similarities = []

    def store(self, state, action, reward, value, log_prob, done, path_idx, similarity):
        self.states.append(state)
        self.actions.append(action)
        self.rewards.append(reward)
        self.values.append(value)
        self.log_probs.append(log_prob)
        self.dones.append(done)
        self.path_indices.append(path_idx)
        self.similarities.append(similarity)

    def compute_advantages(self):
        rewards = np.array(self.rewards)
        values = np.array(self.values)
        dones = np.array(self.dones)

        advantages = np.zeros_like(rewards)
        returns = np.zeros_like(rewards)
        last_advantage = 0

        for t in reversed(range(len(rewards))):
            if t == len(rewards) - 1:
                next_value = 0
            else:
                next_value = values[t + 1]

            delta = rewards[t] + EXPERIMENT_CONFIG['GAMMA'] * next_value * (1 - dones[t]) - values[t]
            advantages[t] = delta + EXPERIMENT_CONFIG['GAMMA'] * EXPERIMENT_CONFIG['GAE_LAMBDA'] * (
                    1 - dones[t]) * last_advantage
            last_advantage = advantages[t]

        returns = advantages + values
        advantages = (advantages - np.mean(advantages)) / (np.std(advantages) + 1e-8)

        return advantages, returns

    def get_batch_iterator(self, advantages, returns):
        indices = np.arange(len(self.states))
        np.random.shuffle(indices)

        batch_size = EXPERIMENT_CONFIG['PPO_BATCH_SIZE']
        for start in range(0, len(self.states), batch_size):
            end = min(start + batch_size, len(self.states))
            batch_indices = indices[start:end]

            yield {
                'states': torch.FloatTensor(np.array([self.states[i] for i in batch_indices])).to(device),
                'actions': torch.FloatTensor(np.array([self.actions[i] for i in batch_indices])).to(device),
                'old_log_probs': torch.FloatTensor(np.array([self.log_probs[i] for i in batch_indices])).to(device),
                'advantages': torch.FloatTensor(advantages[batch_indices]).to(device),
                'returns': torch.FloatTensor(returns[batch_indices]).to(device)
            }

    def get_top_k_per_path(self, num_paths, k=20):
        path_samples = {i: [] for i in range(num_paths)}

        for idx in range(len(self.states)):
            path_idx = self.path_indices[idx]
            similarity = self.similarities[idx]
            state = self.states[idx]
            path_samples[path_idx].append((idx, similarity, state))

        top_k_results = {}
        for path_idx in range(num_paths):
            samples = path_samples[path_idx]
            if len(samples) == 0:
                top_k_results[path_idx] = []
                continue

            samples.sort(key=lambda x: x[1], reverse=True)
            top_k = samples[:k]

            top_k_results[path_idx] = []
            for sample in top_k:
                normalized_state = sample[2]
                original_state = denormalize_state(normalized_state)
                original_state_int = np.round(original_state).astype(int)

                triggered = execute_Tr(*original_state_int)
                top_k_results[path_idx].append({
                    'state': original_state_int,
                    'similarity': sample[1],
                    'triggered': triggered
                })

        return top_k_results

    def clear(self):
        self.states.clear()
        self.actions.clear()
        self.rewards.clear()
        self.values.clear()
        self.log_probs.clear()
        self.dones.clear()
        self.path_indices.clear()
        self.similarities.clear()

    def __len__(self):
        return len(self.states)

# === PPO ===
class PPOAgent:
    def __init__(self, state_dim=3, action_dim=3):
        self.state_dim = state_dim
        self.action_dim = action_dim

        self.actor = PPOActor(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.actor_optimizer = optim.Adam(self.actor.parameters(), lr=EXPERIMENT_CONFIG['ACTOR_LR'])

        self.critic = PPOCritic(state_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_optimizer = optim.Adam(self.critic.parameters(), lr=EXPERIMENT_CONFIG['CRITIC_LR'])

        self.buffer = PPOBuffer()
        self.update_count = 0

    def get_action(self, state):
        min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
        max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
        normalized_state = (state - (min_vals + max_vals) / 2) / ((max_vals - min_vals) / 2)
        state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)

        with torch.no_grad():
            action, log_prob = self.actor.get_action_and_log_prob(state_tensor)
            value = self.critic(state_tensor)

        action = action.cpu().numpy()[0]
        log_prob = log_prob.cpu().item()
        value = value.cpu().item()

        return action, log_prob, value

    def store_experience(self, state, action, reward, value, log_prob, done, path_idx, similarity):
        min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
        max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
        normalized_state = (state - (min_vals + max_vals) / 2) / ((max_vals - min_vals) / 2)
        self.buffer.store(normalized_state, action, reward, value, log_prob, done, path_idx, similarity)

    def update(self):
        if len(self.buffer) == 0:
            return

        advantages, returns = self.buffer.compute_advantages()

        for epoch in range(EXPERIMENT_CONFIG['PPO_EPOCHS']):
            for batch in self.buffer.get_batch_iterator(advantages, returns):
                new_log_probs = self.actor.get_log_prob(batch['states'], batch['actions'])
                ratio = torch.exp(new_log_probs - batch['old_log_probs'])

                surr1 = ratio * batch['advantages']
                surr2 = torch.clamp(ratio, 1 - EXPERIMENT_CONFIG['CLIP_EPSILON'],
                                    1 + EXPERIMENT_CONFIG['CLIP_EPSILON']) * batch['advantages']
                actor_loss = -torch.min(surr1, surr2).mean()

                self.actor_optimizer.zero_grad()
                actor_loss.backward()
                torch.nn.utils.clip_grad_norm_(self.actor.parameters(), 0.5)
                self.actor_optimizer.step()

                new_values = self.critic(batch['states'])
                critic_loss = F.mse_loss(new_values, batch['returns'])

                self.critic_optimizer.zero_grad()
                critic_loss.backward()
                torch.nn.utils.clip_grad_norm_(self.critic.parameters(), 0.5)
                self.critic_optimizer.step()

        self.update_count += 1
        self.buffer.clear()

        if self.update_count % 2 == 0:
            print(f"  -> PPO completed (Run {self.update_count})")

# === Metric ===
def calculate_run_performance(run_idx, ppo_results, training_time, total_steps, update_count, agent):
    """ runMetric"""
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    # Similarity
    all_similarities = []

    # Metric
    total_samples = 0
    all_rewards = []

    for path_idx in range(num_paths):
        samples = ppo_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']

            all_rewards.append(reward)
            all_similarities.append(similarity)
            total_samples += 1

    # 1. Total Reward
    total_reward = sum(all_rewards)

    # 2. Average Reward
    average_reward = total_reward / total_samples if total_samples > 0 else 0

    # 5. Convergence (Average Similarity)
    convergence = np.mean(all_similarities) if all_similarities else 0

    # 12. Environment Adaptability (1/std of similarity)
    environment_adaptability = 1 / (np.std(all_similarities) + 1e-8) if len(all_similarities) > 1 else 0

    # 13. Generalization Ability (Average Similarity)
    generalization_ability = convergence

    # 15. Computational Efficiency (steps/second)
    computational_efficiency = total_steps / training_time if training_time > 0 else 0

    # 16. Policy Update Frequency
    policy_update_frequency = update_count / training_time if training_time > 0 else 0

    # Similarity statistics
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0

    return {
        'Run': run_idx + 1,

        # Metric
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),

        # Similarity
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }

# === Excel ===
def export_to_excel(all_ppo_results, all_performance_data, target_paths, output_path="PPO_20_run.xlsx"):
    """20 run PPO Excel"""
    print("\nExcel...")

    # 
    all_ppo_summary_data = []
    all_ppo_detailed_data = []

    #  run
    for run_idx, (ppo_results, performance_data) in enumerate(zip(all_ppo_results, all_performance_data)):
        # ===== Sheet1: PPO Path  =====
        ppo_summary_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = ppo_results[path_idx]

            if len(samples) == 0:
                ppo_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue

            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = 'Yes' if perfect_count > 0 else 'No'

            ppo_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })

        all_ppo_summary_data.extend(ppo_summary_data)

        # ===== Sheet2: PPO Detailed Sample Data =====
        ppo_detailed_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = ppo_results[path_idx]

            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']

                ppo_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Sample ID': sample_idx + 1,
                    'X': state[0],
                    'Y': state[1],
                    'Z': state[2],
                    'Similarity': similarity,
                    'Triggered Branches': len(triggered),
                    'Triggered Branch List': ', '.join(map(str, sorted(triggered))),
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })

        all_ppo_detailed_data.extend(ppo_detailed_data)

    # Excel
    ppo_summary_df = pd.DataFrame(all_ppo_summary_data)
    ppo_detailed_df = pd.DataFrame(all_ppo_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet1: PPO Path 
        ppo_summary_df.to_excel(writer, sheet_name='PPO Path Summary', index=False)

        # Sheet2: PPO Detailed Sample Data
        ppo_detailed_df.to_excel(writer, sheet_name='PPO Detailed Sample Data', index=False)

        # Sheet3: Metric - 
        selected_columns = [
            'Run',
            'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df_selected = performance_df[selected_columns]
        performance_df_selected.to_excel(writer, sheet_name='Performance Metrics', index=False)

        # 
        workbook = writer.book

        # 
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        perfect_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # 

        # === Sheet1 ===
        ws1 = writer.sheets['PPO Path Summary']
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 
        for row_idx in range(2, ws1.max_row + 1):
            if ws1.cell(row_idx, 9).value == 'Yes':  # Column 9 is Perfect Coverage
                for col_idx in range(1, ws1.max_column + 1):
                    ws1.cell(row_idx, col_idx).fill = perfect_fill

        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 18
        ws1.column_dimensions['D'].width = 15
        ws1.column_dimensions['E'].width = 20
        ws1.column_dimensions['F'].width = 18
        ws1.column_dimensions['G'].width = 18
        ws1.column_dimensions['H'].width = 18
        ws1.column_dimensions['I'].width = 18
        ws1.column_dimensions['J'].width = 50

        # === Sheet2 ===
        ws2 = writer.sheets['PPO Detailed Sample Data']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 15
        ws2.column_dimensions['H'].width = 18
        ws2.column_dimensions['I'].width = 40
        ws2.column_dimensions['J'].width = 50

        # === Sheet3 ===
        ws3 = writer.sheets['Performance Metrics']
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        for col in columns:
            ws3.column_dimensions[col].width = 22

    print(f"Excel: {output_path}")
    print(f"  - Sheet1: PPO Path Summary ({len(all_ppo_summary_data)})")
    print(f"  - Sheet2: PPO Detailed Sample Data ({len(all_ppo_detailed_data)})")
    print(f"  - Sheet3: Performance Metrics ({len(all_performance_data)})")

# === PPO ===
def train_ppo_workflow():
    print("=" * 80)
    print("PPO")
    print("Similarity:  / target paths")
    print("=" * 80)

    agent = PPOAgent()
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n: Path {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']}")
    path_samples = {}
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']

    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            # 分别生成 X, Y, Z 的随机整数
            state = np.array([
                np.random.randint(min_vals[0], max_vals[0] + 1),
                np.random.randint(min_vals[1], max_vals[1] + 1),
                np.random.randint(min_vals[2], max_vals[2] + 1)
            ], dtype=np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  Path {path_idx + 1}/{num_paths}: {len(samples)} ")

    # 
    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size

    print(f"\n: {batch_size},{EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(f": {num_batches} /Path x {num_paths} Path = {num_batches * num_paths} ")
    print("-" * 80)

    global_buffer = PPOBuffer()

    for batch_idx in range(num_batches):
        print(f"\n {batch_idx + 1}/{num_batches}")

        for path_idx in range(num_paths):
            target_path = target_paths[path_idx]
            batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]

            batch_rewards = []
            batch_similarities = []

            for sample_idx, initial_state in enumerate(batch_samples):
                state = initial_state.copy()
                episode_reward = 0
                final_similarity = 0

                for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                    action, log_prob, value = agent.get_action(state)

                    next_state = state + action
                    next_state = clip_state(next_state)

                    triggered = execute_Tr(*next_state)
                    reward = unified_reward_function(triggered, target_path)
                    similarity = coverage_similarity(triggered, target_path)

                    done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)

                    agent.store_experience(state, action, reward, value, log_prob, done, path_idx, similarity)
                    normalized_state = (state - (min_vals + max_vals) / 2) / ((max_vals - min_vals) / 2)
                    global_buffer.store(
                        normalized_state,
                        action, reward, value, log_prob, done, path_idx, similarity
                    )

                    state = next_state
                    episode_reward += reward
                    final_similarity = similarity
                    total_steps += 1

                batch_rewards.append(episode_reward)
                batch_similarities.append(final_similarity)

            avg_reward = np.mean(batch_rewards)
            avg_similarity = np.mean(batch_similarities)
            print(f"  Path {path_idx + 1}: ={avg_reward:.2f}, Average Similarity={avg_similarity:.4f}")

        print(f"\n  PPO...")
        agent.update()
        print(f"  : {len(global_buffer)}")

    training_time = time.time() - start_time

    print("\n" + "=" * 80)
    print(f"PPO completed! Total elapsed time: {training_time:.2f} seconds, : {total_steps}")
    print(f": {len(global_buffer)}")
    print(f"PPO: {agent.update_count}")
    print("=" * 80)

    # Top-K
    print(f"\nPath Similarity Maximum {EXPERIMENT_CONFIG['TOP_K_SAMPLES']}...")
    ppo_top_k_results = global_buffer.get_top_k_per_path(num_paths, EXPERIMENT_CONFIG['TOP_K_SAMPLES'])

    return agent, ppo_top_k_results, training_time, total_steps, agent.update_count

# ===  ===
def main():
    print("\n" + "=" * 80)
    print("PPO - 20 run")
    print("Metric")
    print("=" * 80)

    all_ppo_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    # 20
    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'=' * 80}")
        print(f"Start run {run_idx + 1}/{EXPERIMENT_CONFIG['NUM_RUNS']} run")
        print(f"{'=' * 80}")

        # PPO
        ppo_agent, ppo_results, training_time, total_steps, update_count = train_ppo_workflow()

        # Metric
        performance_data = calculate_run_performance(
            run_idx, ppo_results, training_time, total_steps, update_count, ppo_agent
        )

        # 
        all_ppo_results.append(ppo_results)
        all_performance_data.append(performance_data)

        print(f"\nRun {run_idx + 1} run completed!")
        print(f"  Total Reward: {performance_data['Total Reward']}")
        print(f"  Average Reward: {performance_data['Average Reward']}")
        print(f"  Average Similarity: {performance_data['Average Similarity']}")

    # Excel(20 run)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"PPO_20_runs_{timestamp}.xlsx"
    export_to_excel(all_ppo_results, all_performance_data, target_paths, output_path)

    # 
    print("\n" + "=" * 80)
    print("20 run")
    print("=" * 80)

    # Metric
    total_rewards = [p['Total Reward'] for p in all_performance_data]
    average_rewards = [p['Average Reward'] for p in all_performance_data]
    convergences = [p['Convergence'] for p in all_performance_data]
    environment_adaptabilities = [p['Environment Adaptability'] for p in all_performance_data]
    generalization_abilities = [p['Generalization Ability'] for p in all_performance_data]
    computational_efficiencies = [p['Computational Efficiency'] for p in all_performance_data]
    policy_update_frequencies = [p['Policy Update Frequency'] for p in all_performance_data]
    avg_similarities = [p['Average Similarity'] for p in all_performance_data]

    print(f"\nTotal Reward Statistics:")
    print(f"  Mean: {np.mean(total_rewards):.2f}")
    print(f"  Standard deviation: {np.std(total_rewards):.2f}")

    print(f"\nAverage Reward Statistics:")
    print(f"  Mean: {np.mean(average_rewards):.4f}")
    print(f"  Standard deviation: {np.std(average_rewards):.4f}")

    print(f"\nConvergence Statistics:")
    print(f"  Mean: {np.mean(convergences):.4f}")
    print(f"  Standard deviation: {np.std(convergences):.4f}")

    print(f"\nEnvironment Adaptability Statistics:")
    print(f"  Mean: {np.mean(environment_adaptabilities):.4f}")
    print(f"  Standard deviation: {np.std(environment_adaptabilities):.4f}")

    print(f"\nGeneralization Ability Statistics:")
    print(f"  Mean: {np.mean(generalization_abilities):.4f}")
    print(f"  Standard deviation: {np.std(generalization_abilities):.4f}")

    print(f"\nComputational Efficiency Statistics:")
    print(f"  Mean: {np.mean(computational_efficiencies):.2f}")
    print(f"  Standard deviation: {np.std(computational_efficiencies):.2f}")

    print(f"\nPolicy Update Frequency Statistics:")
    print(f"  Mean: {np.mean(policy_update_frequencies):.4f}")
    print(f"  Standard deviation: {np.std(policy_update_frequencies):.4f}")

    print("\n" + "=" * 80)
    print(f" {EXPERIMENT_CONFIG['NUM_RUNS']} completed!")
    print("=" * 80)

if __name__ == "__main__":
    main()