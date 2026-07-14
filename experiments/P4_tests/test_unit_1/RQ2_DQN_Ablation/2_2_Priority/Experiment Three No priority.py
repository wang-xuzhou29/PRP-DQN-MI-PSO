import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# X, Y, Z 范围
X_MIN = 2
X_MAX = 100
Y_MIN = 1
Y_MAX = 105
Z_MIN = 1
Z_MAX = 110


# === 状态归一化/反归一化 ===
def normalize_state(state):
    """将原始状态归一化到(0-1)"""
    x, y, z = state
    norm_x = (x - X_MIN) / (X_MAX - X_MIN)
    norm_y = (y - Y_MIN) / (Y_MAX - Y_MIN)
    norm_z = (z - Z_MIN) / (Z_MAX - Z_MIN)
    return (norm_x, norm_y, norm_z)


def denormalize_state(norm_state):
    """将归一化状态反归一化到原始范围"""
    norm_x, norm_y, norm_z = norm_state
    x = int(round(norm_x * (X_MAX - X_MIN) + X_MIN))
    y = int(round(norm_y * (Y_MAX - Y_MIN) + Y_MIN))
    z = int(round(norm_z * (Z_MAX - Z_MIN) + Z_MIN))
    x = max(X_MIN, min(X_MAX, x))
    y = max(Y_MIN, min(Y_MAX, y))
    z = max(Z_MIN, min(Z_MAX, z))
    return (x, y, z)


def is_valid_state(state):
    """检查状态是否在有效范围内"""
    x, y, z = state
    return (X_MIN <= x <= X_MAX and
            Y_MIN <= y <= Y_MAX and
            Z_MIN <= z <= Z_MAX)


def clip_state(state):
    """将状态裁剪到有效范围内"""
    x, y, z = state
    return (
        max(X_MIN, min(X_MAX, x)),
        max(Y_MIN, min(Y_MAX, y)),
        max(Z_MIN, min(Z_MAX, z))
    )


# === 增强型标准指标收集器 ===
class EnhancedStandardMetricsCollector:
    def __init__(self, experiment_name="Enhanced_Standard_DQN"):
        self.experiment_name = experiment_name
        self.start_time = None
        self.end_time = None
        self.total_reward = 0
        self.td_errors = []
        self.final_output_similarities = []
        self.action_improvements = []
        self.total_memory_usage = 0
        self.memory_check_count = 0
        self.step_count = 0
        self.episode_rewards = []
        self.episode_similarities = []
        self.episode_td_errors = []
        self.episode_epsilon_values = []
        self.episode_memory_usage = []
        self.similar_paths_performance = []
        self.isolated_paths_performance = []
        self.milestone_data = {}
        self.convergence_window = 20
        self.convergence_threshold = 0.02
        self.convergence_detected_episode = None
        self.sample_efficiency_data = []
        self.performance_milestones = [0.6, 0.7, 0.75, 0.8]
        self.learning_curve_characteristics = {}
        self.early_vs_late_performance = {}
        self.xyz_coordinate_stats = {
            'x_values': [], 'y_values': [], 'z_values': [],
            'x_distribution': {}, 'y_distribution': {}, 'z_distribution': {}
        }

    def reset(self):
        self.start_time = None
        self.end_time = None
        self.total_reward = 0
        self.td_errors = []
        self.final_output_similarities = []
        self.action_improvements = []
        self.total_memory_usage = 0
        self.memory_check_count = 0
        self.step_count = 0
        self.episode_rewards = []
        self.episode_similarities = []
        self.episode_td_errors = []
        self.episode_epsilon_values = []
        self.episode_memory_usage = []
        self.similar_paths_performance = []
        self.isolated_paths_performance = []
        self.milestone_data = {}
        self.convergence_detected_episode = None
        self.sample_efficiency_data = []
        self.learning_curve_characteristics = {}
        self.early_vs_late_performance = {}
        self.xyz_coordinate_stats = {
            'x_values': [], 'y_values': [], 'z_values': [],
            'x_distribution': {}, 'y_distribution': {}, 'z_distribution': {}
        }

    def start_training(self):
        self.start_time = time.time()

    def end_training(self):
        self.end_time = time.time()

    def record_xyz_coordinates(self, x_val, y_val, z_val):
        self.xyz_coordinate_stats['x_values'].append(x_val)
        self.xyz_coordinate_stats['y_values'].append(y_val)
        self.xyz_coordinate_stats['z_values'].append(z_val)
        self.xyz_coordinate_stats['x_distribution'][x_val] = self.xyz_coordinate_stats['x_distribution'].get(x_val, 0) + 1
        self.xyz_coordinate_stats['y_distribution'][y_val] = self.xyz_coordinate_stats['y_distribution'].get(y_val, 0) + 1
        self.xyz_coordinate_stats['z_distribution'][z_val] = self.xyz_coordinate_stats['z_distribution'].get(z_val, 0) + 1

    def record_step_metrics(self, reward, td_error, triggered, target_path, x_coord=None, y_coord=None, z_coord=None):
        self.step_count += 1
        self.total_reward += reward
        self.td_errors.append(td_error)
        if x_coord is not None and y_coord is not None and z_coord is not None:
            self.record_xyz_coordinates(x_coord, y_coord, z_coord)
        process = psutil.Process(os.getpid())
        current_memory = process.memory_info().rss / 1024 / 1024
        self.total_memory_usage += current_memory
        self.memory_check_count += 1

    def record_episode_metrics(self, episode, episode_reward, avg_similarity, avg_td_error, epsilon, path_group="similar"):
        self.episode_rewards.append(episode_reward)
        self.episode_similarities.append(avg_similarity)
        self.episode_td_errors.append(avg_td_error)
        self.episode_epsilon_values.append(epsilon)
        process = psutil.Process(os.getpid())
        current_memory = process.memory_info().rss / 1024 / 1024
        self.episode_memory_usage.append(current_memory)
        if path_group == "similar":
            self.similar_paths_performance.append({
                'episode': episode, 'reward': episode_reward,
                'similarity': avg_similarity, 'td_error': avg_td_error
            })
        else:
            self.isolated_paths_performance.append({
                'episode': episode, 'reward': episode_reward,
                'similarity': avg_similarity, 'td_error': avg_td_error
            })
        if episode in [50, 100, 150, 200, 250, 300, 400, 450, 500]:
            self.milestone_data[episode] = {
                'avg_reward': np.mean(self.episode_rewards[-10:]) if len(self.episode_rewards) >= 10 else episode_reward,
                'avg_similarity': avg_similarity,
                'avg_td_error': avg_td_error,
                'epsilon': epsilon,
                'memory_usage': current_memory,
                'total_steps': self.step_count
            }
        self._check_convergence(episode)
        self._check_performance_milestones(episode, avg_similarity)

    def _check_convergence(self, episode):
        if len(self.episode_similarities) >= self.convergence_window and self.convergence_detected_episode is None:
            recent_similarities = self.episode_similarities[-self.convergence_window:]
            if np.std(recent_similarities) < self.convergence_threshold:
                self.convergence_detected_episode = episode

    def _check_performance_milestones(self, episode, similarity):
        for milestone in self.performance_milestones:
            if similarity >= milestone and not any(data[1] == milestone for data in self.sample_efficiency_data):
                self.sample_efficiency_data.append((episode, milestone, self.step_count))

    def record_final_output_sample(self, triggered, target_path):
        if len(triggered | target_path) > 0:
            similarity = len(triggered & target_path) / len(triggered | target_path)
        else:
            similarity = 0.0
        self.final_output_similarities.append(similarity)

    def record_action_improvement(self, current_reward, prev_reward):
        if prev_reward is not None:
            improvement = current_reward - prev_reward
            self.action_improvements.append(1 if improvement > 0 else 0)


# 全局指标收集器
enhanced_standard_metrics = EnhancedStandardMetricsCollector("Enhanced_Standard_DQN_No_Priority")


# === 奖励函数 ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward


# === 核心变异分支函数（返回triggered集合）===
def execute_Tr(x, y, z):
    triggered = set()

    # --- 分支 1-14 ---
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 1) > 120): triggered.add(1)
    if ((y * z) / (x + 1) > 100) != ((y * y) / (x + 1) > 100): triggered.add(2)
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 8) > 100): triggered.add(3)
    if ((y * z) / (x + 1) > 100) != ((y + z) / (x + 1) > 100): triggered.add(4)
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x * 3 + 1) > 100): triggered.add(5)
    if ((y * z) / (x + 1) > 100) != ((z * z) / (x + 1) > 100): triggered.add(6)
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 1) > 80): triggered.add(7)
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + y) > 100): triggered.add(8)
    if ((y * z) / (x + 1) > 100) != ((y * 10) / (x + 1) > 100): triggered.add(9)
    if ((y * z) / (x + 1) > 100) != ((x * z) / (x + 1) > 100): triggered.add(10)
    if ((y * z) / (x + 1) > 100) != ((y * z * 2) / (x + 1) > 100): triggered.add(11)
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 5) > 100): triggered.add(12)
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 1) > 200): triggered.add(13)
    if ((y * z) / (x + 1) > 100) != ((5 * z) / (x + 1) > 100): triggered.add(14)

    # --- 分支 15-22 ---
    if ((z - x) < 0.3 * y) != ((z * 1.2 - x) < 0.3 * y): triggered.add(15)
    if ((z - x) < 0.3 * y) != ((z - x * 1.1) < 0.3 * y): triggered.add(16)
    if ((z - x) < 0.3 * y) != ((z - x) < 0.5 * y): triggered.add(17)
    if ((z - x) < 0.3 * y) != ((z - x) < 0.3 * z): triggered.add(18)
    if ((z - x) < 0.3 * y) != ((z - x) < 0.3 * y * y): triggered.add(19)
    if ((z - x) < 0.3 * y) != ((z - y) < 0.3 * y): triggered.add(20)
    if ((z - x) < 0.3 * y) != ((z - x) < 0.8 * y): triggered.add(21)
    if ((z - x) < 0.3 * y) != ((z - x) < 0.3 * x): triggered.add(22)

    # --- 分支 23-36 ---
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 3): triggered.add(23)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 * y ** 3) < z ** 4): triggered.add(24)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2 + y ** 3) < z ** 2): triggered.add(25)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 2) < z ** 2): triggered.add(26)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * 3) < (z ** 2)): triggered.add(27)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * x ** 3) < z ** 2): triggered.add(28)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 1 + y ** 3) < z ** 2): triggered.add(29)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 1) < z ** 2): triggered.add(30)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x * 3 + y ** 3) < z ** 2): triggered.add(31)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x * 3 + y ** 3) < z ** 2): triggered.add(32)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * 3) < z ** 2): triggered.add(33)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** x + y ** 3) < z ** 2): triggered.add(34)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2 + y ** 3) < z ** 2): triggered.add(35)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 1.5): triggered.add(36)

    # --- 分支 37-54 ---
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 or (y / (z + 0.001) < 0.2)): triggered.add(37)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y % (z + 0.001) < 0.2)): triggered.add(38)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y / (z - 0.001) < 0.4)): triggered.add(39)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.3)): triggered.add(40)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.1)): triggered.add(41)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 8 and (y / (z + 0.001) < 0.2)): triggered.add(42)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 2 and (y / (z + 0.001) < 0.2)): triggered.add(43)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 1 and (y / (z + 0.001) < 0.2)): triggered.add(44)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 10 and (y / (z + 0.001) < 0.2)): triggered.add(45)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x * y / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.2)): triggered.add(46)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x * x / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.2)): triggered.add(47)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y * x / (z + 0.001) < 0.2)): triggered.add(48)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y * y / (z + 0.001) < 0.2)): triggered.add(49)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y * z / (z + 0.001) < 0.2)): triggered.add(50)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.5)): triggered.add(51)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y - (z + 0.001) < 0.2)): triggered.add(52)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x + (y + 0.001)) > 5 and (y / (z + 0.001) < 0.2)): triggered.add(53)
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x - (y + 0.001)) > 5 and (y / (z + 0.001) < 0.2)): triggered.add(54)

    # --- 分支 55-75 ---
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 20): triggered.add(55)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 12.9): triggered.add(56)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(20 - z) < 5): triggered.add(57)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - 80) > 10 and abs(x - z) < 5): triggered.add(58)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x + z) < 5): triggered.add(59)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 7): triggered.add(60)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 15 and abs(10 - z) < 10): triggered.add(61)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - 8) < 5): triggered.add(62)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - 2) < 5): triggered.add(63)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x * 4 - z) < 5): triggered.add(64)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x * z) < 5): triggered.add(65)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x + z) < 5): triggered.add(66)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x * y - z) < 5): triggered.add(67)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 2): triggered.add(68)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - y) < 5): triggered.add(69)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 15): triggered.add(70)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x * 2 - z) < 5): triggered.add(71)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(y - z) < 5): triggered.add(72)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x * z - z) < 5): triggered.add(73)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z * 2) < 5): triggered.add(74)
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x * x - z) < 5): triggered.add(75)

    # --- 分支 76-85 ---
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x * x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(76)
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x * y > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(77)
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x * z > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(78)
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x > 90 or x < 5) and (y * y > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(79)
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x > 90 or x < 5) and (y * z > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(80)
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x > 90 or x < 5) and (y * x > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(81)
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x * 10 > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(82)
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x * 15 > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(83)
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x > 50 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): triggered.add(84)
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != ((x > 90 or x < 5) and (y > 80 or y < 3) and (z * 40 > 75 or z < 2)): triggered.add(85)

    # --- 分支 86-100 ---
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * x) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): triggered.add(86)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * 60) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): triggered.add(87)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): triggered.add(88)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((70 * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): triggered.add(89)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) + (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): triggered.add(90)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 70 and x ** 2 + y ** 2 > z ** 2): triggered.add(91)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 80 and x ** 2 + y ** 2 > z ** 2): triggered.add(92)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 1.5 + y ** 2 > z ** 2): triggered.add(93)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2.2): triggered.add(94)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2.5): triggered.add(95)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 1.5 > z ** 2): triggered.add(96)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 80 and x ** 2 + y ** 2 > z ** 2): triggered.add(97)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 65 and x ** 2 + y ** 2 > z ** 2): triggered.add(98)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 1.2 + y ** 2 > z ** 2): triggered.add(99)
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 1.2 > z ** 2): triggered.add(100)

    # --- 分支 101-105 ---
    if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.7 > (x + y) / 2 and x * y * z > 1000): triggered.add(101)
    if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.5 > (x + y) / 6 and x * y * z > 1000): triggered.add(102)
    if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.8 > (x + y) / 2 and x * y * z > 1000): triggered.add(103)
    if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.5 > (x + x) / 2 and x * y * z > 1000): triggered.add(104)
    if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.5 > (y + y) / 2 and x * y * z > 1000): triggered.add(105)

    return triggered


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0


# === 路径相似度计算 ===
def compute_path_similarity_matrix(paths):
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix


targetPaths = [
    {15, 20, 29, 31, 32, 38, 41, 42, 45, 48, 49, 50, 57, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 79, 80, 81, 86, 87, 90},
    {6, 10, 15, 20, 23, 29, 31, 32, 38, 41, 42, 45, 48, 49, 50, 57, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 79, 80, 81},
    {1, 3, 4, 5, 6, 8, 9, 10, 13, 14, 57, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 76, 77, 78, 82, 83, 84, 94, 95, 100},
    {16, 17, 18, 19, 21, 22, 23, 29, 31, 32, 38, 41, 45, 48, 49, 50, 57, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 105},
    {6, 7, 10, 11, 15, 57, 58, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 79, 80, 81, 92, 93, 94, 95, 96, 97, 99, 100},
    {1, 2, 3, 4, 5, 8, 9, 10, 13, 14, 16, 17, 18, 19, 20, 21, 23, 76, 77, 78, 82, 83, 84, 87, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100},
    {11, 15, 20, 57, 59, 61, 62, 63, 64, 65, 66, 67, 69, 71, 72, 73, 74, 75, 79, 80, 81, 88, 89, 91, 92, 93, 94, 95, 97, 98, 99, 100},
    {1, 3, 4, 5, 6, 8, 9, 10, 12, 13, 14, 18, 22, 57, 58, 59, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 86, 87, 91, 92, 97, 98},
    {16, 18, 19, 21, 22, 23, 24, 25, 29, 31, 32, 35, 38, 48, 50, 57, 59, 61, 62, 63, 64, 65, 66, 67, 69, 71, 72, 73, 74, 75, 105},
    {15, 20, 29, 31, 32, 38, 41, 42, 45, 48, 49, 50, 59, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 87, 90, 102, 105},
    {1, 2, 3, 4, 5, 8, 9, 10, 12, 13, 14, 19, 23, 24, 25, 26, 27, 29, 30, 31, 32, 33, 35, 37, 46, 47, 53, 101, 102, 103},
    {3, 4, 5, 6, 8, 10, 12, 13, 14, 18, 22, 57, 59, 62, 64, 65, 66, 67, 68, 69, 72, 73, 74, 85, 86, 87, 91, 92, 97, 98},
    {15, 20, 23, 29, 31, 32, 37, 43, 44, 46, 47, 53, 54, 57, 59, 61, 62, 63, 64, 65, 66, 67, 69, 71, 72, 73, 74, 75},
    {6, 10, 16, 18, 19, 21, 22, 23, 25, 29, 31, 32, 35, 38, 45, 48, 49, 50, 55, 56, 60, 70, 79, 80, 81, 105},
    {3, 4, 5, 6, 8, 9, 10, 12, 13, 14, 17, 19, 20, 21, 27, 30, 33, 55, 57, 61, 63, 64, 70, 88, 89, 90, 104},
    {2, 3, 4, 5, 8, 9, 10, 14, 19, 23, 24, 26, 27, 28, 30, 33, 62, 63, 67, 101, 102, 103, 104},
    {1, 2, 3, 4, 5, 8, 9, 10, 12, 13, 14, 19, 28, 34, 36, 37, 46, 47, 53, 101, 102, 103},
    {15, 20, 29, 31, 32, 37, 39, 40, 51, 52, 55, 56, 70, 79, 80, 81, 86, 87, 90}
]


# === 路径分组函数 ===
def group_paths_by_similarity(paths):
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.mean(avg_sim_scores)
    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]
    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)
    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group


# === 标准经验回放(无优先级) ===
class StandardExperienceReplay:
    def __init__(self, capacity=10000):
        self.capacity = capacity
        self.buffer = deque(maxlen=self.capacity)

    def append(self, experience):
        self.buffer.append(experience[:5])

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return [], [], None
        batch_indices = np.random.choice(len(self.buffer), batch_size, replace=False)
        batch = [self.buffer[idx] for idx in batch_indices]
        return batch, batch_indices, None

    def __len__(self):
        return len(self.buffer)

    def get_high_reward_samples(self, target_path, num_samples=20):
        if len(self.buffer) == 0:
            return []
        samples_with_recalculated_scores = []
        for experience in self.buffer:
            norm_state_tensor = experience[0]
            norm_state = tuple(norm_state_tensor.cpu().numpy().flatten())
            state_tuple = denormalize_state(norm_state)
            triggered = execute_Tr(state_tuple[0], state_tuple[1], state_tuple[2])
            new_reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_recalculated_scores.append((state_tuple, new_reward, sim, triggered))
        samples_with_recalculated_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_recalculated_scores[:num_samples]


def load_path_data(file_path):
    path_data = []
    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines[2:]:
            parts = line.strip().split("\t")
            state = tuple(map(int, parts[0].split()))
            path_data.append(state)
    return path_data


# === DQN网络 ===
class DQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, 128)
        self.fc2 = nn.Linear(128, 64)
        self.fc3 = nn.Linear(64, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)


# === 标准DQN Agent(无优先级) ===
class StandardDQNAgent:
    def __init__(self, state_dim, action_dim, replay_buffer, gamma=0.99, epsilon=1.0,
                 epsilon_decay=0.995, epsilon_min=0.1, learning_rate=0.001):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer
        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        delta_values = [10, 5, 3, 2, 1, -1, -2, -3, -5, -10]
        dim = action_idx // 10
        delta_idx = action_idx % 10
        delta = delta_values[delta_idx]
        if dim == 0:
            return (delta, 0, 0)
        elif dim == 1:
            return (0, delta, 0)
        elif dim == 2:
            return (0, 0, delta)

    def act(self, state):
        if random.random() < self.epsilon:
            return random.randrange(self.action_dim)
        norm_state = normalize_state(state)
        state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)
        return torch.argmax(q_values, dim=1).item()

    def store_transition(self, state, action, reward, next_state, done):
        norm_state = normalize_state(state)
        norm_next_state = normalize_state(next_state)
        norm_state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
        norm_next_state_tensor = torch.tensor(norm_next_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(norm_state_tensor)
            next_q_values = self.target_model(norm_next_state_tensor)
            max_next_q_values = next_q_values.max(1)[0]
            target_q_values = reward + (self.gamma * max_next_q_values * (1 - done))
            td_error = torch.abs(q_values[0][action] - target_q_values).item()
        self.replay_buffer.append((norm_state_tensor, action, reward, norm_next_state_tensor, done))
        return td_error

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return
        batch, batch_indices, _ = self.replay_buffer.sample(batch_size)
        states, actions, rewards, next_states, dones = zip(*batch)
        states = torch.tensor(np.array([s.cpu().numpy().flatten() for s in states]), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array([ns.cpu().numpy().flatten() for ns in next_states]), dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)
        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))
        loss = nn.MSELoss()(current_q_values, target_q_values)
        self.optimizer.zero_grad()
        loss.backward()
        self.optimizer.step()
        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())


# === 样本生成函数 ===
def generate_samples_for_similar_paths(similar_group_indices, num_total=2000, top_k=200):
    def jaccard_similarity_local(a, b):
        if not a and not b: return 1.0
        return len(a & b) / len(a | b) if a | b else 0.0

    def compute_robustness(state, path):
        base = execute_Tr(state[0], state[1], state[2])
        if not base: return 0.0
        rob, neighbors = 0.0, 0
        for dx in [-1, 0, 1]:
            for dy in [-1, 0, 1]:
                for dz in [-1, 0, 1]:
                    if dx == dy == dz == 0: continue
                    neighbor_state = (state[0] + dx, state[1] + dy, state[2] + dz)
                    if not is_valid_state(neighbor_state): continue
                    neighbor = clip_state(neighbor_state)
                    n_trig = execute_Tr(neighbor[0], neighbor[1], neighbor[2])
                    if not n_trig: continue
                    rob += jaccard_similarity_local(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"enhanced_standard_path{path_id}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Enhanced Standard Path {path_id}\n")
            f.write("x y z\tScore\tSimilarity\tLengthDiff\tRobustness\n")
            for s in samples:
                x, y, z = s[0]
                f.write(f"{x} {y} {z}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_enhanced_standard"
    for path_idx in similar_group_indices:
        path = targetPaths[path_idx]
        samples = []
        attempts = 0
        while len(samples) < top_k and attempts < num_total * 5:
            attempts += 1
            state = (
                np.random.randint(X_MIN, X_MAX + 1),
                np.random.randint(Y_MIN, Y_MAX + 1),
                np.random.randint(Z_MIN, Z_MAX + 1)
            )
            triggered = execute_Tr(state[0], state[1], state[2])
            if not triggered: continue
            sim = jaccard_similarity_local(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            score = 0.55 * sim + 0.25 * len_diff + 0.2 * rob
            samples.append((state, score, sim, len_diff, rob))
        if samples:
            samples.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples[:top_k], base_dir=base_dir)


# === 相似路径训练 ===
def enhanced_standard_generate_and_train_for_similar_paths(agent, similar_group, path_documents, episodes=500,
                                                           batch_size=32, steps_per_test=5, replay_times=10,
                                                           is_isolated=False):
    trained_paths = set()
    update_target_every = 100
    global_steps = 0

    for episode in range(episodes):
        episode_reward = 0
        episode_similarities = []
        episode_td_errors = []

        for path_idx in similar_group:
            if path_idx in trained_paths:
                continue
            file_path = os.path.join(path_documents,
                                     f"enhanced_standard_path{path_idx + 1}{'_isolated' if is_isolated else ''}.txt")
            path_data = load_path_data(file_path)
            target_path = targetPaths[path_idx]

            BATCH_SIZE = 50
            N_SAMPLES = 200
            N_STEPS = 10
            REPLAY_TIMES = 3

            for batch_start in range(0, N_SAMPLES, BATCH_SIZE):
                batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)
                for test_data in range(batch_start, batch_end):
                    if test_data >= len(path_data):
                        break
                    step_count = 0
                    state = path_data[test_data]
                    prev_state = None
                    prev_triggered = None
                    prev_reward = None

                    for step in range(N_STEPS):
                        legal_actions = []
                        for a in range(agent.action_dim):
                            dx, dy, dz = agent.decode_action(a)
                            cand_next = (state[0] + dx, state[1] + dy, state[2] + dz)
                            if is_valid_state(cand_next):
                                legal_actions.append(a)
                        if not legal_actions:
                            break
                        if random.random() < agent.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            norm_state = normalize_state(state)
                            state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        dx, dy, dz = agent.decode_action(action)
                        next_state = clip_state((state[0] + dx, state[1] + dy, state[2] + dz))
                        triggered = execute_Tr(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                        done = (step_count == N_STEPS - 1)
                        td_error = agent.store_transition(state, action, reward, next_state, done)
                        enhanced_standard_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                      next_state[0], next_state[1], next_state[2])
                        episode_similarities.append(jaccard_similarity(triggered, target_path))
                        episode_td_errors.append(td_error)
                        if prev_reward is not None:
                            enhanced_standard_metrics.record_action_improvement(reward, prev_reward)
                        prev_state = state
                        prev_triggered = triggered
                        prev_reward = reward
                        state = next_state
                        step_count += 1
                        episode_reward += reward
                        global_steps += 1
                        if global_steps % update_target_every == 0:
                            agent.update_target_model()

                for _ in range(REPLAY_TIMES):
                    if len(agent.replay_buffer) >= batch_size:
                        agent.train(batch_size)
            trained_paths.add(path_idx)

        avg_similarity = np.mean(episode_similarities) if episode_similarities else 0
        avg_td_error = np.mean(episode_td_errors) if episode_td_errors else 0
        enhanced_standard_metrics.record_episode_metrics(episode, episode_reward, avg_similarity, avg_td_error,
                                                         agent.epsilon, "similar")
        if episode % 10 == 0:
            agent.update_target_model()
        if len(trained_paths) == len(similar_group):
            break
    return agent


def generate_samples_for_isolated_paths_standard(agent_similar, isolated_group_indices, num_total=2000, top_k=200):
    def compute_q_value(state, agent):
        norm_state = normalize_state(state)
        state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = agent.model(state_tensor)
        return q_values.max().item()

    def compute_robustness(state, path):
        base = execute_Tr(state[0], state[1], state[2])
        if not base: return 0.0
        rob, neighbors = 0.0, 0
        for dx in [-1, 0, 1]:
            for dy in [-1, 0, 1]:
                for dz in [-1, 0, 1]:
                    if dx == dy == dz == 0: continue
                    neighbor_state = (state[0] + dx, state[1] + dy, state[2] + dz)
                    if not is_valid_state(neighbor_state): continue
                    neighbor = clip_state(neighbor_state)
                    n_trig = execute_Tr(neighbor[0], neighbor[1], neighbor[2])
                    if not n_trig: continue
                    rob += jaccard_similarity(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"enhanced_standard_path{path_id}_isolated.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Enhanced Standard Isolated Path {path_id}\n")
            f.write("x y z\tScore\tSimilarity\tLengthDiff\tRobustness\tQ_complement\n")
            for s in samples:
                x, y, z = s[0]
                f.write(f"{x} {y} {z}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\t{s[5]:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_enhanced_standard"
    for path_idx in isolated_group_indices:
        path = targetPaths[path_idx]
        samples_raw = []
        attempts = 0
        print(f"生成隔离路径 {path_idx + 1} 样本...")
        while len(samples_raw) < num_total and attempts < num_total * 5:
            attempts += 1
            state = (
                np.random.randint(X_MIN, X_MAX + 1),
                np.random.randint(Y_MIN, Y_MAX + 1),
                np.random.randint(Z_MIN, Z_MAX + 1)
            )
            triggered = execute_Tr(state[0], state[1], state[2])
            if not triggered: continue
            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            q_value = compute_q_value(state, agent_similar)
            samples_raw.append((state, sim, len_diff, rob, q_value))

        if not samples_raw:
            print(f"路径 {path_idx + 1}: 未生成有效样本")
            continue

        q_values_list = [s[4] for s in samples_raw]
        q_min = min(q_values_list)
        q_max = max(q_values_list)
        print(f"路径 {path_idx + 1}: 生成了 {len(samples_raw)} 个候选样本, Q值范围: [{q_min:.4f}, {q_max:.4f}]")

        samples_final = []
        for state, sim, len_diff, rob, q_value in samples_raw:
            if q_max - q_min > 1e-6:
                q_normalized = (q_value - q_min) / (q_max - q_min)
            else:
                q_normalized = 0.5
            q_complement = 1.0 - q_normalized
            score = 0.28 * sim + 0.1 * len_diff + 0.19 * rob + 0.43 * q_complement
            samples_final.append((state, score, sim, len_diff, rob, q_complement))

        if samples_final:
            samples_final.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples_final[:top_k], base_dir=base_dir)
            print(f"路径 {path_idx + 1}: 保存了 {min(top_k, len(samples_final))} 个最终样本")


# === 隔离路径训练 ===
def enhanced_standard_generate_and_train_for_isolated_paths(agent_similar, agent_isolated, similar_group,
                                                            isolated_group, path_documents, episodes=500, batch_size=32,
                                                            is_isolated=True):
    trained_paths = set()
    update_target_every = 100
    global_steps = 0
    stage1_samples_pool = {}

    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_similar.replay_buffer.get_high_reward_samples(target_path, num_samples=100)
        stage1_samples_pool[path_idx] = high_reward_samples

    for episode in range(episodes):
        episode_reward = 0
        episode_similarities = []
        episode_td_errors = []

        for path_idx in isolated_group:
            if path_idx in trained_paths:
                continue
            file_path = os.path.join(path_documents, f"enhanced_standard_path{path_idx + 1}_isolated.txt")
            stage2_path_data = load_path_data(file_path)
            stage1_samples = stage1_samples_pool.get(path_idx, [])
            target_path = targetPaths[path_idx]

            BATCH_SIZE = 50
            N_SAMPLES_STAGE2 = min(140, len(stage2_path_data))
            N_SAMPLES_STAGE1 = min(60, len(stage1_samples))
            N_STEPS = 10
            REPLAY_TIMES = 3

            for batch_start in range(0, N_SAMPLES_STAGE2, BATCH_SIZE):
                batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES_STAGE2)
                for test_data_idx in range(batch_start, batch_end):
                    if test_data_idx >= len(stage2_path_data):
                        break
                    step_count = 0
                    state = stage2_path_data[test_data_idx]
                    prev_state = None
                    prev_triggered = None
                    prev_reward = None

                    for step in range(N_STEPS):
                        legal_actions = []
                        for a in range(agent_isolated.action_dim):
                            dx, dy, dz = agent_isolated.decode_action(a)
                            cand_next = (state[0] + dx, state[1] + dy, state[2] + dz)
                            if is_valid_state(cand_next):
                                legal_actions.append(a)
                        if not legal_actions:
                            break
                        if random.random() < agent_isolated.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            norm_state = normalize_state(state)
                            state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent_isolated.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        dx, dy, dz = agent_isolated.decode_action(action)
                        next_state = clip_state((state[0] + dx, state[1] + dy, state[2] + dz))
                        triggered = execute_Tr(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                        if target_path.issubset(triggered):
                            reward += 2.0
                        done = (step_count == N_STEPS - 1)
                        td_error = agent_isolated.store_transition(state, action, reward, next_state, done)
                        enhanced_standard_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                      next_state[0], next_state[1], next_state[2])
                        episode_similarities.append(jaccard_similarity(triggered, target_path))
                        episode_td_errors.append(td_error)
                        if prev_reward is not None:
                            enhanced_standard_metrics.record_action_improvement(reward, prev_reward)
                        prev_state = state
                        prev_triggered = triggered
                        prev_reward = reward
                        state = next_state
                        step_count += 1
                        episode_reward += reward
                        global_steps += 1
                        if global_steps % update_target_every == 0:
                            agent_isolated.update_target_model()

            if stage1_samples:
                for sample_idx in range(N_SAMPLES_STAGE1):
                    if sample_idx >= len(stage1_samples):
                        break
                    stage1_state_tuple, _, _, _ = stage1_samples[sample_idx]
                    step_count = 0
                    state = stage1_state_tuple
                    prev_state = None
                    prev_triggered = None
                    prev_reward = None

                    for step in range(N_STEPS):
                        legal_actions = []
                        for a in range(agent_isolated.action_dim):
                            dx, dy, dz = agent_isolated.decode_action(a)
                            cand_next = (state[0] + dx, state[1] + dy, state[2] + dz)
                            if is_valid_state(cand_next):
                                legal_actions.append(a)
                        if not legal_actions:
                            break
                        if random.random() < agent_isolated.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            norm_state = normalize_state(state)
                            state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent_isolated.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        dx, dy, dz = agent_isolated.decode_action(action)
                        next_state = clip_state((state[0] + dx, state[1] + dy, state[2] + dz))
                        triggered = execute_Tr(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                        reward *= 0.8
                        done = (step_count == N_STEPS - 1)
                        td_error = agent_isolated.store_transition(state, action, reward, next_state, done)
                        enhanced_standard_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                      next_state[0], next_state[1], next_state[2])
                        episode_similarities.append(jaccard_similarity(triggered, target_path))
                        episode_td_errors.append(td_error)
                        if prev_reward is not None:
                            enhanced_standard_metrics.record_action_improvement(reward, prev_reward)
                        prev_state = state
                        prev_triggered = triggered
                        prev_reward = reward
                        state = next_state
                        step_count += 1
                        episode_reward += reward
                        global_steps += 1
                        if global_steps % update_target_every == 0:
                            agent_isolated.update_target_model()

            for replay_round in range(REPLAY_TIMES):
                if len(agent_isolated.replay_buffer) >= batch_size:
                    agent_isolated.train(batch_size)
            trained_paths.add(path_idx)

        avg_similarity = np.mean(episode_similarities) if episode_similarities else 0
        avg_td_error = np.mean(episode_td_errors) if episode_td_errors else 0
        enhanced_standard_metrics.record_episode_metrics(episode, episode_reward, avg_similarity, avg_td_error,
                                                         agent_isolated.epsilon, "isolated")
        if episode % 10 == 0:
            agent_isolated.update_target_model()
        if len(trained_paths) == len(isolated_group):
            break
    return agent_isolated


# === Excel报告生成 ===
def append_metrics_to_combined_excel(metrics_collector, agent_similar, agent_isolated, similar_group, isolated_group,
                                     targetPaths, filepath, run_number):
    training_time = metrics_collector.end_time - metrics_collector.start_time if metrics_collector.end_time else 0
    avg_memory = metrics_collector.total_memory_usage / metrics_collector.memory_check_count if metrics_collector.memory_check_count > 0 else 0
    avg_similarity = np.mean(metrics_collector.final_output_similarities) if metrics_collector.final_output_similarities else 0
    avg_td_error = np.mean(metrics_collector.td_errors) if metrics_collector.td_errors else 0
    action_improve_rate = np.mean(metrics_collector.action_improvements) if metrics_collector.action_improvements else 0

    performance_row = {
        'Run': run_number,
        'Average Similarity': f"{avg_similarity:.4f}",
        'TD Error': f"{avg_td_error:.4f}",
        'Action Improve Rate': f"{action_improve_rate:.4f}",
        'Memory(MB)': f"{avg_memory:.2f}"
    }

    sample_rows = []
    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_similar.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, reward, sim, triggered in high_reward_samples:
            sample_rows.append({
                'Run': run_number, 'Group Type': 'Similar', 'Path ID': path_idx + 1,
                'X': state_tuple[0], 'Y': state_tuple[1], 'Z': state_tuple[2],
                'Similarity': f"{sim:.4f}", 'Reward': f"{reward:.2f}",
                'Triggered Count': len(triggered), 'Target Count': len(target_path),
                'Triggered Rules': str(sorted(triggered)), 'Target Rules': str(sorted(target_path))
            })
    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, reward, sim, triggered in high_reward_samples:
            sample_rows.append({
                'Run': run_number, 'Group Type': 'Isolated', 'Path ID': path_idx + 1,
                'X': state_tuple[0], 'Y': state_tuple[1], 'Z': state_tuple[2],
                'Similarity': f"{sim:.4f}", 'Reward': f"{reward:.2f}",
                'Triggered Count': len(triggered), 'Target Count': len(target_path),
                'Triggered Rules': str(sorted(triggered)), 'Target Rules': str(sorted(target_path))
            })

    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    if os.path.exists(filepath):
        try:
            df_performance = pd.read_excel(filepath, sheet_name='性能指标')
            df_performance = pd.concat([df_performance, pd.DataFrame([performance_row])], ignore_index=True)
        except:
            df_performance = pd.DataFrame([performance_row])
        try:
            df_samples = pd.read_excel(filepath, sheet_name='最终样本')
            df_samples = pd.concat([df_samples, pd.DataFrame(sample_rows)], ignore_index=True)
        except:
            df_samples = pd.DataFrame(sample_rows)
    else:
        df_performance = pd.DataFrame([performance_row])
        df_samples = pd.DataFrame(sample_rows)

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df_performance.to_excel(writer, sheet_name='性能指标', index=False)
        df_samples.to_excel(writer, sheet_name='最终样本', index=False)
        workbook = writer.book
        ws_performance = writer.sheets['性能指标']
        ws_performance.column_dimensions['A'].width = 15
        for col in ['B', 'C', 'D', 'E']:
            ws_performance.column_dimensions[col].width = 20
        header_font = Font(bold=True, size=11)
        for cell in ws_performance[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in ws_performance.iter_rows(min_row=2, max_row=ws_performance.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        ws_samples = writer.sheets['最终样本']
        column_widths = {'A': 12, 'B': 15, 'C': 12, 'D': 10, 'E': 10, 'F': 10,
                         'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 40, 'L': 40}
        for col, width in column_widths.items():
            ws_samples.column_dimensions[col].width = width
        for cell in ws_samples[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in ws_samples.iter_rows(min_row=2, max_row=ws_samples.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    print(f"第 {run_number} 次运行结果已保存到: {filepath}")


# === 单次实验运行(含文件检查和自动重新生成) ===
def run_single_experiment(run_number, results_save_dir):
    print(f"\n{'=' * 80}")
    print(f"开始第 {run_number} 次运行")
    print(f"{'=' * 80}\n")

    enhanced_standard_metrics.reset()
    enhanced_standard_metrics.start_training()

    model_path_similar = os.path.join(results_save_dir, f"run{run_number}_similar.pth")
    model_path_isolated = os.path.join(results_save_dir, f"run{run_number}_isolated.pth")
    path_documents = r"D:\Experiment\CNN\DQNNEW\path_samples_enhanced_standard"

    similar_group, isolated_group = group_paths_by_similarity(targetPaths)

    # 检查并生成缺失的样本文件
    def check_and_generate_samples():
        all_files_exist = True
        for path_idx in similar_group:
            file_path = os.path.join(path_documents, f"enhanced_standard_path{path_idx + 1}.txt")
            if not os.path.exists(file_path):
                all_files_exist = False
                print(f"  [缺失] enhanced_standard_path{path_idx + 1}.txt")
        for path_idx in isolated_group:
            file_path = os.path.join(path_documents, f"enhanced_standard_path{path_idx + 1}_isolated.txt")
            if not os.path.exists(file_path):
                all_files_exist = False
                print(f"  [缺失] enhanced_standard_path{path_idx + 1}_isolated.txt")
        return all_files_exist

    print("正在检查样本文件...")
    files_exist = check_and_generate_samples()

    if not files_exist:
        print("\n检测到缺失的样本文件，正在自动重新生成...")
        os.makedirs(path_documents, exist_ok=True)
        print("\n[1/2] 生成相似路径组样本...")
        generate_samples_for_similar_paths(similar_group, num_total=2000, top_k=200)
        print("相似路径组样本生成完成！")

        temp_agent = None
        prev_run = max(1, run_number - 1)
        prev_model_path = os.path.join(results_save_dir, f"run{prev_run}_similar.pth")
        if os.path.exists(prev_model_path):
            try:
                temp_replay_buffer = StandardExperienceReplay(capacity=10000)
                temp_agent = StandardDQNAgent(3, 30, temp_replay_buffer)
                checkpoint = torch.load(prev_model_path)
                temp_agent.model.load_state_dict(checkpoint['model_state_dict'])
                temp_agent.target_model.load_state_dict(checkpoint['model_state_dict'])
                print(f"\n成功加载第 {prev_run} 次运行的模型用于样本生成")
            except Exception as e:
                print(f"\n加载前一次模型失败: {e}")
                temp_agent = None
        if temp_agent is None:
            print("使用新初始化的agent生成隔离路径样本...")
            temp_replay_buffer = StandardExperienceReplay(capacity=10000)
            temp_agent = StandardDQNAgent(3, 30, temp_replay_buffer)

        print("\n[2/2] 生成隔离路径组样本...")
        generate_samples_for_isolated_paths_standard(temp_agent, isolated_group, num_total=2000, top_k=200)
        print("隔离路径组样本生成完成！")
        print("\n再次检查样本文件...")
        if check_and_generate_samples():
            print("所有样本文件已成功生成！")
        else:
            print("警告：部分文件仍然缺失，但将继续运行...")

    print("\n样本文件检查完成，开始训练...\n")

    # 阶段1: 相似路径训练
    replay_buffer = StandardExperienceReplay(capacity=10000)
    state_dim = 3
    action_dim = 30
    agent = StandardDQNAgent(state_dim, action_dim, replay_buffer)
    agent = enhanced_standard_generate_and_train_for_similar_paths(agent, similar_group, path_documents, episodes=500,
                                                                   batch_size=32, is_isolated=False)
    os.makedirs(os.path.dirname(model_path_similar), exist_ok=True)
    torch.save({
        'model_state_dict': agent.model.state_dict(),
        'optimizer_state_dict': agent.optimizer.state_dict(),
        'epsilon': agent.epsilon
    }, model_path_similar)

    # 阶段2: 隔离路径训练
    isolated_replay_buffer = StandardExperienceReplay(capacity=15000)
    agent_isolated = StandardDQNAgent(state_dim, action_dim, isolated_replay_buffer)
    try:
        checkpoint = torch.load(model_path_similar)
        agent_isolated.model.load_state_dict(checkpoint['model_state_dict'])
        agent_isolated.target_model.load_state_dict(checkpoint['model_state_dict'])
        agent_isolated.epsilon = checkpoint.get('epsilon', 0.5)
        print("成功加载相似路径模型用于隔离路径训练")
    except Exception as e:
        print(f"无法加载相似路径模型: {e}")

    agent_isolated = enhanced_standard_generate_and_train_for_isolated_paths(
        agent_similar=agent, agent_isolated=agent_isolated,
        similar_group=similar_group, isolated_group=isolated_group,
        path_documents=path_documents, episodes=500, batch_size=32, is_isolated=True
    )
    os.makedirs(os.path.dirname(model_path_isolated), exist_ok=True)
    torch.save({
        'model_state_dict': agent_isolated.model.state_dict(),
        'optimizer_state_dict': agent_isolated.optimizer.state_dict(),
        'epsilon': agent_isolated.epsilon
    }, model_path_isolated)

    enhanced_standard_metrics.end_training()

    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, _, sim, triggered in high_reward_samples:
            enhanced_standard_metrics.record_final_output_sample(triggered, target_path)
    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, _, sim, triggered in high_reward_samples:
            enhanced_standard_metrics.record_final_output_sample(triggered, target_path)

    combined_excel_path = os.path.join(results_save_dir, "综合实验结果.xlsx")
    append_metrics_to_combined_excel(
        metrics_collector=enhanced_standard_metrics,
        agent_similar=agent, agent_isolated=agent_isolated,
        similar_group=similar_group, isolated_group=isolated_group,
        targetPaths=targetPaths, filepath=combined_excel_path, run_number=run_number
    )

    avg_similarity = np.mean(enhanced_standard_metrics.final_output_similarities)
    print(f"\n第 {run_number} 次运行完成:")
    print(f"  平均相似度: {avg_similarity:.4f}")
    print(f"  总步数: {enhanced_standard_metrics.step_count}")


if __name__ == "__main__":
    results_save_dir = r"D:\Experiment\CNN\_"
    os.makedirs(results_save_dir, exist_ok=True)
    path_documents = r"D:\Experiment\CNN\DQNNEW\path_samples_enhanced_standard"
    os.makedirs(path_documents, exist_ok=True)

    NUM_RUNS = 20
    print("=" * 80)
    print(f"开始 {NUM_RUNS} 次标准DQN实验(无优先级)")
    print(f"参数范围: X[{X_MIN}, {X_MAX}], Y[{Y_MIN}, {Y_MAX}], Z[{Z_MIN}, {Z_MAX}]")
    print("=" * 80)

    for run in range(1, NUM_RUNS + 1):
        try:
            run_single_experiment(run, results_save_dir)
        except Exception as e:
            print(f"\n第 {run} 次运行出错: {str(e)}")
            import traceback
            traceback.print_exc()
            continue

    print("\n" + "=" * 80)
    print(f" {NUM_RUNS} 次运行全部完成")
    print(f"结果保存目录: {results_save_dir}")
    print("=" * 80)