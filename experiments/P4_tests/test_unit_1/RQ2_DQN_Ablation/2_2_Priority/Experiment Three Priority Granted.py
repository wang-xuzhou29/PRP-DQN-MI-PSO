import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# === 设备设置 ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === 状态空间范围 ===
dx_min, dx_max = 2, 100  # dx范围
dy_min, dy_max = 1, 105  # dy范围
dz_min, dz_max = 1, 110  # dz范围


# === 状态归一化函数 ===
def normalize_state(state):
    """
    将状态归一化到[0,1]范围
    state: (dx, dy, dz)
    """
    dx, dy, dz = state
    normalized_dx = (dx - dx_min) / (dx_max - dx_min)
    normalized_dy = (dy - dy_min) / (dy_max - dy_min)
    normalized_dz = (dz - dz_min) / (dz_max - dz_min)
    return (normalized_dx, normalized_dy, normalized_dz)


def denormalize_state(normalized_state):
    """
    将归一化状态还原为原始范围
    """
    norm_dx, norm_dy, norm_dz = normalized_state
    dx = int(norm_dx * (dx_max - dx_min) + dx_min)
    dy = int(norm_dy * (dy_max - dy_min) + dy_min)
    dz = int(norm_dz * (dz_max - dz_min) + dz_min)
    return (dx, dy, dz)


def is_valid_state(state):
    """检查状态是否在有效范围内"""
    dx, dy, dz = state
    return (dx_min <= dx <= dx_max and
            dy_min <= dy <= dy_max and
            dz_min <= dz <= dz_max)


def clip_state(state):
    """将状态裁剪到有效范围内"""
    dx, dy, dz = state
    return (
        max(dx_min, min(dx_max, dx)),
        max(dy_min, min(dy_max, dy)),
        max(dz_min, min(dz_max, dz))
    )


# === 指标收集器 ===
class PrioritizedMetricsCollector:
    def __init__(self, experiment_name="Prioritized_DQN"):
        self.experiment_name = experiment_name
        self.start_time = None
        self.end_time = None

        # 步骤级指标
        self.total_reward = 0
        self.td_errors = []
        self.final_output_similarities = []
        self.action_improvements = []
        self.total_memory_usage = 0
        self.memory_check_count = 0
        self.step_count = 0

        # 回合级指标
        self.episode_rewards = []
        self.episode_similarities = []
        self.episode_td_errors = []
        self.episode_epsilon_values = []
        self.episode_memory_usage = []

        # 优先级相关指标
        self.priority_statistics = []
        self.importance_weights = []
        self.high_priority_samples_ratio = []
        self.priority_distribution_stats = []

        # 路径分组性能
        self.similar_paths_performance = []
        self.isolated_paths_performance = []

        # 里程碑数据
        self.milestone_data = {}

        # 收敛检测
        self.convergence_window = 20
        self.convergence_threshold = 0.02
        self.convergence_detected_episode = None

        # 样本效率
        self.sample_efficiency_data = []
        self.performance_milestones = [0.6, 0.7, 0.75, 0.8]

        # 学习曲线特征
        self.learning_curve_characteristics = {}
        self.early_vs_late_performance = {}

    def reset(self):
        """重置所有指标，用于多次运行"""
        self.start_time = None
        self.end_time = None
        self.total_reward = 0
        self.td_errors = []
        self.final_output_similarities = []
        self.action_improvements = []
        self.total_memory_usage = 0
        self.memory_check_count = 0
        self.step_count = 0
        self.episode_rewards = []
        self.episode_similarities = []
        self.episode_td_errors = []
        self.episode_epsilon_values = []
        self.episode_memory_usage = []
        self.priority_statistics = []
        self.importance_weights = []
        self.high_priority_samples_ratio = []
        self.priority_distribution_stats = []
        self.similar_paths_performance = []
        self.isolated_paths_performance = []
        self.milestone_data = {}
        self.convergence_detected_episode = None
        self.sample_efficiency_data = []
        self.learning_curve_characteristics = {}
        self.early_vs_late_performance = {}

    def start_training(self):
        self.start_time = time.time()

    def end_training(self):
        self.end_time = time.time()

    def record_step_metrics(self, reward, td_error, triggered, target_path, priority=None, is_weight=None):
        """记录步骤级指标"""
        self.step_count += 1
        self.total_reward += reward
        self.td_errors.append(td_error)

        if priority is not None:
            self.priority_statistics.append(priority)
        if is_weight is not None:
            self.importance_weights.append(is_weight)

        process = psutil.Process(os.getpid())
        current_memory = process.memory_info().rss / 1024 / 1024
        self.total_memory_usage += current_memory
        self.memory_check_count += 1

    def record_episode_metrics(self, episode, episode_reward, avg_similarity, avg_td_error, epsilon,
                               path_group="similar", priority_stats=None):
        """记录回合级指标"""
        self.episode_rewards.append(episode_reward)
        self.episode_similarities.append(avg_similarity)
        self.episode_td_errors.append(avg_td_error)
        self.episode_epsilon_values.append(epsilon)

        process = psutil.Process(os.getpid())
        current_memory = process.memory_info().rss / 1024 / 1024
        self.episode_memory_usage.append(current_memory)

        if priority_stats:
            self.priority_distribution_stats.append({
                'episode': episode,
                'mean_priority': priority_stats.get('mean_priority', 0),
                'max_priority': priority_stats.get('max_priority', 0),
                'min_priority': priority_stats.get('min_priority', 0),
                'high_priority_ratio': priority_stats.get('high_priority_ratio', 0)
            })

        if path_group == "similar":
            self.similar_paths_performance.append({
                'episode': episode,
                'reward': episode_reward,
                'similarity': avg_similarity,
                'td_error': avg_td_error
            })
        else:
            self.isolated_paths_performance.append({
                'episode': episode,
                'reward': episode_reward,
                'similarity': avg_similarity,
                'td_error': avg_td_error
            })

        if episode in [50, 100, 150, 200, 250, 300, 400, 450, 500]:
            self.milestone_data[episode] = {
                'avg_reward': np.mean(self.episode_rewards[-10:]) if len(
                    self.episode_rewards) >= 10 else episode_reward,
                'avg_similarity': avg_similarity,
                'avg_td_error': avg_td_error,
                'epsilon': epsilon,
                'memory_usage': current_memory,
                'total_steps': self.step_count,
                'priority_stats': priority_stats
            }

        self._check_convergence(episode)
        self._check_performance_milestones(episode, avg_similarity)

    def _check_convergence(self, episode):
        """检查是否收敛"""
        if len(self.episode_similarities) >= self.convergence_window and self.convergence_detected_episode is None:
            recent_similarities = self.episode_similarities[-self.convergence_window:]
            if np.std(recent_similarities) < self.convergence_threshold:
                self.convergence_detected_episode = episode

    def _check_performance_milestones(self, episode, similarity):
        """检查性能里程碑"""
        for milestone in self.performance_milestones:
            if similarity >= milestone and not any(data[1] == milestone for data in self.sample_efficiency_data):
                self.sample_efficiency_data.append((episode, milestone, self.step_count))

    def record_final_output_sample(self, triggered, target_path):
        """记录最终输出样本的相似度"""
        if len(triggered | target_path) > 0:
            similarity = len(triggered & target_path) / len(triggered | target_path)
        else:
            similarity = 0.0
        self.final_output_similarities.append(similarity)

    def record_action_improvement(self, current_reward, prev_reward):
        """记录动作改进情况"""
        if prev_reward is not None:
            improvement = current_reward - prev_reward
            self.action_improvements.append(1 if improvement > 0 else 0)


# 全局指标收集器
prioritized_metrics = PrioritizedMetricsCollector("Prioritized_DQN_Enhanced")


# === 奖励函数 ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward


def execute_Tr(x, y, z):
    """执行测试需求，返回触发的分支集合"""
    # 初始化分支覆盖数组
    b = [0] * 105  # 根据分支数量调整大小

    # --- 分支 1-14 ---
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 1) > 120): b[0] = 1
    if ((y * z) / (x + 1) > 100) != ((y * y) / (x + 1) > 100): b[1] = 2
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 8) > 100): b[2] = 3
    if ((y * z) / (x + 1) > 100) != ((y + z) / (x + 1) > 100): b[3] = 4
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x * 3 + 1) > 100): b[4] = 5
    if ((y * z) / (x + 1) > 100) != ((z * z) / (x + 1) > 100): b[5] = 6
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 1) > 80): b[6] = 7
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + y) > 100): b[7] = 8
    if ((y * z) / (x + 1) > 100) != ((y * 10) / (x + 1) > 100): b[8] = 9
    if ((y * z) / (x + 1) > 100) != ((x * z) / (x + 1) > 100): b[9] = 10
    if ((y * z) / (x + 1) > 100) != ((y * z * 2) / (x + 1) > 100): b[10] = 11
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 5) > 100): b[11] = 12
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 1) > 200): b[12] = 13
    if ((y * z) / (x + 1) > 100) != ((5 * z) / (x + 1) > 100): b[13] = 14

    # --- 分支 15-22 ---
    if ((z - x) < 0.3 * y) != ((z * 1.2 - x) < 0.3 * y): b[14] = 15
    if ((z - x) < 0.3 * y) != ((z - x * 1.1) < 0.3 * y): b[15] = 16
    if ((z - x) < 0.3 * y) != ((z - x) < 0.5 * y): b[16] = 17
    if ((z - x) < 0.3 * y) != ((z - x) < 0.3 * z): b[17] = 18
    if ((z - x) < 0.3 * y) != ((z - x) < 0.3 * y * y): b[18] = 19
    if ((z - x) < 0.3 * y) != ((z - y) < 0.3 * y): b[19] = 20
    if ((z - x) < 0.3 * y) != ((z - x) < 0.8 * y): b[20] = 21
    if ((z - x) < 0.3 * y) != ((z - x) < 0.3 * x): b[21] = 22

    # --- 分支 23-36 ---
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 3): b[22] = 23
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 * y ** 3) < z ** 4): b[23] = 24
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2 + y ** 3) < z ** 2): b[24] = 25
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 2) < z ** 2): b[25] = 26
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * 3) < (z ** 2)): b[26] = 27
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * x ** 3) < z ** 2): b[27] = 28
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 1 + y ** 3) < z ** 2): b[28] = 29
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 1) < z ** 2): b[29] = 30
    if ((x ** 3 + y ** 3) < z ** 2) != ((x * 3 + y ** 3) < z ** 2): b[30] = 31
    if ((x ** 3 + y ** 3) < z ** 2) != ((x * 3 + y ** 3) < z ** 2): b[31] = 32
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * 3) < z ** 2): b[32] = 33
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** x + y ** 3) < z ** 2): b[33] = 34
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2 + y ** 3) < z ** 2): b[34] = 35
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 1.5): b[35] = 36

    # --- 分支 37-54 ---
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 or (y / (z + 0.001) < 0.2)): b[
        36] = 37
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y % (z + 0.001) < 0.2)): b[
        37] = 38
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y / (z - 0.001) < 0.4)): b[
        38] = 39
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.3)): b[
        39] = 40
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.1)): b[
        40] = 41
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 8 and (y / (z + 0.001) < 0.2)): b[
        41] = 42
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 2 and (y / (z + 0.001) < 0.2)): b[
        42] = 43
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 1 and (y / (z + 0.001) < 0.2)): b[
        43] = 44
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 10 and (y / (z + 0.001) < 0.2)): b[
        44] = 45
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x * y / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.2)):
        b[45] = 46
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x * x / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.2)):
        b[46] = 47
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y * x / (z + 0.001) < 0.2)):
        b[47] = 48
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y * y / (z + 0.001) < 0.2)):
        b[48] = 49
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y * z / (z + 0.001) < 0.2)):
        b[49] = 50
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.5)): b[
        50] = 51
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x / (y + 0.001)) > 5 and (y - (z + 0.001) < 0.2)): b[
        51] = 52
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x + (y + 0.001)) > 5 and (y / (z + 0.001) < 0.2)): b[
        52] = 53
    if ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2) != ((x - (y + 0.001)) > 5 and (y / (z + 0.001) < 0.2)): b[
        53] = 54

    # --- 分支 55-75 ---
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 20): b[54] = 55
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 12.9): b[55] = 56
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(20 - z) < 5): b[56] = 57
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - 80) > 10 and abs(x - z) < 5): b[57] = 58
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x + z) < 5): b[58] = 59
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 7): b[59] = 60
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 15 and abs(10 - z) < 10): b[60] = 61
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x - 8) < 5): b[61] = 62
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x - 2) < 5): b[62] = 63
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x * 4 - z) < 5): b[63] = 64
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x * z) < 5): b[64] = 65
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x + z) < 5): b[65] = 66
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x * y - z) < 5): b[66] = 67
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 2): b[67] = 68
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x - y) < 5): b[68] = 69
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 15): b[69] = 70
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x * 2 - z) < 5): b[70] = 71
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(y - z) < 5): b[71] = 72
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x * z - z) < 5): b[72] = 73
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z * 2) < 5): b[73] = 74
    if (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5) != (
            abs(x - y) > 10 and abs(y - z) > 10 and abs(x * x - z) < 5): b[74] = 75

    # --- 分支 76-85 ---
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != (
            (x * x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): b[75] = 76
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != (
            (x * y > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): b[76] = 77
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != (
            (x * z > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): b[77] = 78
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != (
            (x > 90 or x < 5) and (y * y > 80 or y < 3) and (z > 75 or z < 2)): b[78] = 79
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != (
            (x > 90 or x < 5) and (y * z > 80 or y < 3) and (z > 75 or z < 2)): b[79] = 80
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != (
            (x > 90 or x < 5) and (y * x > 80 or y < 3) and (z > 75 or z < 2)): b[80] = 81
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != (
            (x * 10 > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): b[81] = 82
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != (
            (x * 15 > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): b[82] = 83
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != (
            (x > 50 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): b[83] = 84
    if ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)) != (
            (x > 90 or x < 5) and (y > 80 or y < 3) and (z * 40 > 75 or z < 2)): b[84] = 85

    # --- 分支 86-100 ---
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * x) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[85] = 86
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * 60) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[86] = 87
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[87] = 88
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((70 * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[88] = 89
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) + (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[89] = 90
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 70 and x ** 2 + y ** 2 > z ** 2):
        b[90] = 91
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 80 and x ** 2 + y ** 2 > z ** 2):
        b[91] = 92
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 1.5 + y ** 2 > z ** 2):
        b[92] = 93
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2.2):
        b[93] = 94
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2.5):
        b[94] = 95
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 1.5 > z ** 2):
        b[95] = 96
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 80 and x ** 2 + y ** 2 > z ** 2):
        b[96] = 97
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 65 and x ** 2 + y ** 2 > z ** 2):
        b[97] = 98
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 1.2 + y ** 2 > z ** 2):
        b[98] = 99
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 1.2 > z ** 2):
        b[99] = 100

    # --- 分支 101-105 ---
    if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.7 > (x + y) / 2 and x * y * z > 1000): b[100] = 101
    if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.5 > (x + y) / 6 and x * y * z > 1000): b[101] = 102
    if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.8 > (x + y) / 2 and x * y * z > 1000): b[102] = 103
    if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.5 > (x + x) / 2 and x * y * z > 1000): b[103] = 104
    if (z ** 0.5 > (x + y) / 2 and x * y * z > 1000) != (z ** 0.5 > (y + y) / 2 and x * y * z > 1000): b[104] = 105

    # 返回触发的分支集合
    triggered = set()
    for i, val in enumerate(b):
        if val > 0:
            triggered.add(i + 1)  # 分支编号从1开始
    return triggered


def jaccard_similarity(set1, set2):
    """计算Jaccard相似度"""
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0


# === 路径相似度计算 ===
def compute_path_similarity_matrix(paths):
    """计算路径相似度矩阵"""
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix


targetPaths = [
    {15, 20, 29, 31, 32, 38, 41, 42, 45, 48, 49, 50, 57, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 79,
     80, 81, 86, 87, 90},  # A1
    {6, 10, 15, 20, 23, 29, 31, 32, 38, 41, 42, 45, 48, 49, 50, 57, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73,
     74, 75, 79, 80, 81},  # A2
    {1, 3, 4, 5, 6, 8, 9, 10, 13, 14, 57, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 76, 77, 78, 82,
     83, 84, 94, 95, 100},  # A3
    {16, 17, 18, 19, 21, 22, 23, 29, 31, 32, 38, 41, 45, 48, 49, 50, 57, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72,
     73, 74, 75, 105},  # A4
    {6, 7, 10, 11, 15, 57, 58, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 79, 80, 81, 92, 93, 94, 95,
     96, 97, 99, 100},  # A5
    {1, 2, 3, 4, 5, 8, 9, 10, 13, 14, 16, 17, 18, 19, 20, 21, 23, 76, 77, 78, 82, 83, 84, 87, 91, 92, 93, 94, 95, 96,
     97, 98, 99, 100},  # A6
    {11, 15, 20, 57, 59, 61, 62, 63, 64, 65, 66, 67, 69, 71, 72, 73, 74, 75, 79, 80, 81, 88, 89, 91, 92, 93, 94, 95, 97,
     98, 99, 100},  # A7
    {1, 3, 4, 5, 6, 8, 9, 10, 12, 13, 14, 18, 22, 57, 58, 59, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 86, 87,
     91, 92, 97, 98},  # A8
    {16, 18, 19, 21, 22, 23, 24, 25, 29, 31, 32, 35, 38, 48, 50, 57, 59, 61, 62, 63, 64, 65, 66, 67, 69, 71, 72, 73, 74,
     75, 105},  # A9
    {15, 20, 29, 31, 32, 38, 41, 42, 45, 48, 49, 50, 59, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 87, 90,
     102, 105},  # A10
    {1, 2, 3, 4, 5, 8, 9, 10, 12, 13, 14, 19, 23, 24, 25, 26, 27, 29, 30, 31, 32, 33, 35, 37, 46, 47, 53, 101, 102,
     103},  # A11
    {3, 4, 5, 6, 8, 10, 12, 13, 14, 18, 22, 57, 59, 62, 64, 65, 66, 67, 68, 69, 72, 73, 74, 85, 86, 87, 91, 92, 97, 98},
    # A12
    {15, 20, 23, 29, 31, 32, 37, 43, 44, 46, 47, 53, 54, 57, 59, 61, 62, 63, 64, 65, 66, 67, 69, 71, 72, 73, 74, 75},
    # A13
    {6, 10, 16, 18, 19, 21, 22, 23, 25, 29, 31, 32, 35, 38, 45, 48, 49, 50, 55, 56, 60, 70, 79, 80, 81, 105},  # A14
    {3, 4, 5, 6, 8, 9, 10, 12, 13, 14, 17, 19, 20, 21, 27, 30, 33, 55, 57, 61, 63, 64, 70, 88, 89, 90, 104},  # A15
    {2, 3, 4, 5, 8, 9, 10, 14, 19, 23, 24, 26, 27, 28, 30, 33, 62, 63, 67, 101, 102, 103, 104},  # A16
    {1, 2, 3, 4, 5, 8, 9, 10, 12, 13, 14, 19, 28, 34, 36, 37, 46, 47, 53, 101, 102, 103},  # A17
    {15, 20, 29, 31, 32, 37, 39, 40, 51, 52, 55, 56, 70, 79, 80, 81, 86, 87, 90}  # A18
]


# === 路径分组 ===
def group_paths_by_similarity(paths):
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.mean(avg_sim_scores)

    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]
    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)

    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group


# === 优先经验回放 ===
class PrioritizedExperienceReplay:
    def __init__(self, capacity=10000, alpha=0.6, beta_start=0.4, beta_frames=100000):
        self.capacity = capacity
        self.alpha = alpha
        self.beta_start = beta_start
        self.beta_frames = beta_frames
        self.frame = 1

        self.buffer = []
        self.priorities = np.zeros(capacity, dtype=np.float32)
        self.pos = 0
        self.size = 0

        self.max_priority = 1.0
        self.min_priority = 1.0

    def beta(self):
        """计算当前的beta值"""
        return min(1.0, self.beta_start + (1.0 - self.beta_start) * self.frame / self.beta_frames)

    def append(self, experience):
        """添加经验到缓冲区"""
        if len(self.buffer) < self.capacity:
            self.buffer.append(experience)
        else:
            self.buffer[self.pos] = experience

        self.priorities[self.pos] = self.max_priority

        self.pos = (self.pos + 1) % self.capacity
        self.size = min(self.size + 1, self.capacity)

    def sample(self, batch_size):
        """从缓冲区采样"""
        if self.size < batch_size:
            return [], [], []

        priorities = self.priorities[:self.size]
        probs = priorities ** self.alpha
        probs /= probs.sum()

        # 不重复采样
        indices = np.random.choice(self.size, batch_size, p=probs, replace=False)

        # 去重：确保批次中没有重复状态
        unique_batch = []
        unique_indices = []
        seen_states = set()

        for idx in indices:
            experience = self.buffer[idx]
            state_tensor = experience[0]
            # 将tensor转换为元组用于去重
            state_tuple = tuple(state_tensor.cpu().numpy().flatten())

            if state_tuple not in seen_states:
                seen_states.add(state_tuple)
                unique_batch.append(experience)
                unique_indices.append(idx)

        # 如果去重后样本不够，补充新样本
        if len(unique_batch) < batch_size:
            remaining_indices = [i for i in range(self.size) if i not in unique_indices]
            if remaining_indices:
                remaining_probs = priorities[remaining_indices] ** self.alpha
                remaining_probs /= remaining_probs.sum()

                needed = batch_size - len(unique_batch)
                additional_indices = np.random.choice(
                    remaining_indices,
                    min(needed, len(remaining_indices)),
                    p=remaining_probs,
                    replace=False
                )

                for idx in additional_indices:
                    experience = self.buffer[idx]
                    state_tensor = experience[0]
                    state_tuple = tuple(state_tensor.cpu().numpy().flatten())

                    if state_tuple not in seen_states:
                        seen_states.add(state_tuple)
                        unique_batch.append(experience)
                        unique_indices.append(idx)

        total = len(self.buffer)
        unique_indices = np.array(unique_indices)
        weights = (total * probs[unique_indices]) ** (-self.beta())
        weights /= weights.max()

        self.frame += 1

        return unique_batch, unique_indices, weights

    def update_priorities(self, indices, priorities):
        """更新经验优先级"""
        for idx, priority in zip(indices, priorities):
            if idx < self.size:
                self.priorities[idx] = priority
                self.max_priority = max(self.max_priority, priority)
                self.min_priority = min(self.min_priority, priority)

    def get_priority_statistics(self):
        """获取优先级统计信息"""
        if self.size == 0:
            return None

        priorities = self.priorities[:self.size]
        mean_priority = np.mean(priorities)
        max_priority = np.max(priorities)
        min_priority = np.min(priorities)

        high_priority_ratio = np.mean(priorities > mean_priority)

        return {
            'mean_priority': mean_priority,
            'max_priority': max_priority,
            'min_priority': min_priority,
            'high_priority_ratio': high_priority_ratio
        }

    def __len__(self):
        return self.size

    def get_high_reward_samples(self, target_path, num_samples=20):
        """获取高奖励样本"""
        if len(self.buffer) == 0:
            return []

        samples_with_recalculated_scores = []
        for experience in self.buffer:
            state_tensor = experience[0]
            state_tuple = tuple(state_tensor.cpu().numpy().flatten().astype(int))
            triggered = execute_Tr(*state_tuple)
            new_reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_recalculated_scores.append((state_tuple, new_reward, sim, triggered))

        samples_with_recalculated_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_recalculated_scores[:num_samples]


def load_path_data(file_path):
    path_data = []
    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines[2:]:
            parts = line.strip().split("\t")
            state = tuple(map(int, parts[0].split()))
            path_data.append(state)
    return path_data


# === DQN网络 ===
class DQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, 128)
        self.fc2 = nn.Linear(128, 64)
        self.fc3 = nn.Linear(64, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)


# === 优先级DQN智能体 ===
class PrioritizedDQNAgent:
    def __init__(self, state_dim, action_dim, replay_buffer, gamma=0.99, epsilon=1.0,
                 epsilon_decay=0.995, epsilon_min=0.1, learning_rate=0.001):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        """
        解码动作索引为状态变化值
        
        动作空间: 30个动作 = 3个维度 x 10个变化值
        - 维度0: dx (1-50)
        - 维度1: dy (1-50)
        - 维度2: dz (1-50)

        每个维度的变化值: 
        - 正向: 35(70%), 25(50%), 10(20%), 5(10%), 2(5%) (共5个正向变化)
        - 负向: -2(5%), -5(10%), -10(20%), -25(50%), -35(70%) (共5个负向变化)
        """
        # 10个变化值：5个正向，5个负向
        delta_values = [35, 25, 10, 5, 2, -2, -5, -10, -25, -35]

        dim = action_idx // 10
        delta_idx = action_idx % 10

        if dim == 0:  # dx
            return (delta_values[delta_idx], 0, 0)
        elif dim == 1:  # dy
            return (0, delta_values[delta_idx], 0)
        elif dim == 2:  # dz
            return (0, 0, delta_values[delta_idx])
        else:
            return (0, 0, 0)

    def act(self, state):
        if random.random() < self.epsilon:
            return random.randrange(self.action_dim)
        # 贪婪动作选择
        normalized_state = normalize_state(state)
        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)
        return torch.argmax(q_values, dim=1).item()

    def store_transition(self, state, action, reward, next_state, done):
        """存储转换并返回TD误差用于优先级更新"""
        # 归一化状态
        normalized_state = normalize_state(state)
        normalized_next_state = normalize_state(next_state)

        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        next_state_tensor = torch.tensor(normalized_next_state, dtype=torch.float32).unsqueeze(0).to(device)

        with torch.no_grad():
            q_values = self.model(state_tensor)
            next_q_values = self.target_model(next_state_tensor)
            max_next_q_values = next_q_values.max(1)[0]
            target_q_values = reward + (self.gamma * max_next_q_values * (1 - done))
            td_error = torch.abs(q_values[0][action] - target_q_values).item()

        self.replay_buffer.append((state_tensor, action, reward, next_state_tensor, done, td_error))
        return td_error

    def train(self, batch_size=32):
        """训练模型（优先级经验回放版本）"""
        if len(self.replay_buffer) < batch_size:
            return

        batch, batch_indices, is_weights = self.replay_buffer.sample(batch_size)
        if not batch:
            return

        states, actions, rewards, next_states, dones, _ = zip(*batch)

        # 转换为tensor并移到设备
        states = torch.tensor(np.array([s.cpu().numpy().flatten() for s in states]), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array([ns.cpu().numpy().flatten() for ns in next_states]),
                                   dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)
        is_weights = torch.tensor(is_weights, dtype=torch.float32).to(device)

        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))

        td_errors = torch.abs(current_q_values - target_q_values)
        loss = (is_weights * (current_q_values - target_q_values) ** 2).mean()

        self.optimizer.zero_grad()
        loss.backward()
        self.optimizer.step()

        new_priorities = td_errors.detach().cpu().numpy() + 1e-6
        self.replay_buffer.update_priorities(batch_indices, new_priorities)

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())


# === 相似路径样本生成 ===
def generate_samples_for_similar_paths(similar_group_indices, num_total=2000, top_k=200):
    def jaccard_similarity_local(a, b):
        if not a and not b:
            return 1.0
        return len(a & b) / len(a | b) if a | b else 0.0

    def compute_robustness(state, path):
        base = execute_Tr(*state)
        if not base:
            return 0.0
        rob, neighbors = 0.0, 0
        for ddx in [-1, 0, 1]:
            for ddy in [-1, 0, 1]:
                for ddz in [-1, 0, 1]:
                    if ddx == ddy == ddz == 0:
                        continue
                    neighbor_state = (state[0] + ddx, state[1] + ddy, state[2] + ddz)
                    if not is_valid_state(neighbor_state):
                        continue
                    neighbor = clip_state(neighbor_state)
                    n_trig = execute_Tr(*neighbor)
                    if not n_trig:
                        continue
                    rob += jaccard_similarity_local(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"prioritized_path{path_id}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Prioritized Path {path_id}\n")
            f.write("dx dy dz\tScore\tSimilarity\tLengthDiff\tRobustness\n")
            for s in samples:
                dx, dy, dz = s[0]
                f.write(f"{dx} {dy} {dz}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_prioritized"
    for path_idx in similar_group_indices:
        path = targetPaths[path_idx]
        samples = []
        attempts = 0
        while len(samples) < top_k and attempts < num_total * 5:
            attempts += 1
            state = (
                np.random.randint(dx_min, dx_max + 1),
                np.random.randint(dy_min, dy_max + 1),
                np.random.randint(dz_min, dz_max + 1)
            )
            triggered = execute_Tr(*state)
            if not triggered:
                continue
            sim = jaccard_similarity_local(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            score = 0.55 * sim + 0.25 * len_diff + 0.2 * rob
            samples.append((state, score, sim, len_diff, rob))
        if samples:
            samples.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples[:top_k], base_dir=base_dir)


# === 相似路径训练 ===
def prioritized_generate_and_train_for_similar_paths(agent, similar_group, path_documents, episodes=500,
                                                     batch_size=32, steps_per_test=3, replay_times=1,
                                                     is_isolated=False):
    trained_paths = set()
    global_replay_count = 0

    for episode in range(episodes):
        episode_reward = 0
        episode_similarities = []
        episode_td_errors = []

        for path_idx in similar_group:
            if path_idx in trained_paths:
                continue

            file_path = os.path.join(path_documents,
                                     f"prioritized_path{path_idx + 1}{'_isolated' if is_isolated else ''}.txt")
            path_data = load_path_data(file_path)
            target_path = targetPaths[path_idx]

            # 训练参数设置
            BATCH_SIZE = 50
            N_SAMPLES = 200
            N_STEPS = 3
            N_BATCHES = 4
            PATH_REPEAT = 5

            for repeat in range(PATH_REPEAT):
                for batch_idx in range(N_BATCHES):
                    batch_start = batch_idx * BATCH_SIZE
                    batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)

                    for test_data_idx in range(batch_start, batch_end):
                        if test_data_idx >= len(path_data):
                            break

                        state = path_data[test_data_idx]
                        prev_state = None
                        prev_triggered = None
                        prev_reward = None

                        for step in range(N_STEPS):
                            legal_actions = []
                            for a in range(agent.action_dim):
                                ddx, ddy, ddz = agent.decode_action(a)
                                cand_next = (state[0] + ddx, state[1] + ddy, state[2] + ddz)
                                if is_valid_state(cand_next):
                                    legal_actions.append(a)
                            if not legal_actions:
                                break

                            if random.random() < agent.epsilon:
                                action = random.choice(legal_actions)
                            else:
                                # 贪婪动作选择
                                normalized_state = normalize_state(state)
                                state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(
                                    device)
                                with torch.no_grad():
                                    q_values = agent.model(state_tensor)[0]
                                action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                            ddx, ddy, ddz = agent.decode_action(action)
                            next_state = clip_state((state[0] + ddx, state[1] + ddy, state[2] + ddz))

                            triggered = execute_Tr(*next_state)
                            reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                            done = (step == N_STEPS - 1)

                            td_error = agent.store_transition(state, action, reward, next_state, done)

                            priority_stats = agent.replay_buffer.get_priority_statistics()
                            current_priority = priority_stats['mean_priority'] if priority_stats else 0

                            prioritized_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                    current_priority, None)

                            episode_similarities.append(jaccard_similarity(triggered, target_path))
                            episode_td_errors.append(td_error)

                            if prev_reward is not None:
                                prioritized_metrics.record_action_improvement(reward, prev_reward)

                            prev_state = state
                            prev_triggered = triggered
                            prev_reward = reward
                            state = next_state
                            episode_reward += reward

                    if len(agent.replay_buffer) >= batch_size:
                        agent.train(batch_size)
                        global_replay_count += 1

                        if global_replay_count % 2 == 0:
                            agent.update_target_model()

            trained_paths.add(path_idx)

        avg_similarity = np.mean(episode_similarities) if episode_similarities else 0
        avg_td_error = np.mean(episode_td_errors) if episode_td_errors else 0
        priority_stats = agent.replay_buffer.get_priority_statistics()

        prioritized_metrics.record_episode_metrics(episode, episode_reward, avg_similarity, avg_td_error,
                                                   agent.epsilon, "similar", priority_stats)

        if len(trained_paths) == len(similar_group):
            break

    return agent


def generate_samples_for_isolated_paths_prioritized(agent_similar, isolated_group_indices, num_total=2000, top_k=200):
    def compute_q_value_normalized_complement(state, agent):
        """计算归一化Q值补数"""
        # 归一化状态
        normalized_state = normalize_state(state)
        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = agent.model(state_tensor)

        all_q_values = q_values[0].cpu().numpy()

        q_min = all_q_values.min()
        q_max = all_q_values.max()

        if q_max - q_min > 1e-6:
            normalized_q = (all_q_values.max() - q_min) / (q_max - q_min)
        else:
            normalized_q = 0.0

        complement_q = 1.0 - normalized_q

        return complement_q

    def compute_robustness(state, path):
        base = execute_Tr(*state)
        if not base:
            return 0.0
        rob, neighbors = 0.0, 0
        for ddx in [-1, 0, 1]:
            for ddy in [-1, 0, 1]:
                for ddz in [-1, 0, 1]:
                    if ddx == ddy == ddz == 0:
                        continue
                    neighbor_state = (state[0] + ddx, state[1] + ddy, state[2] + ddz)
                    if not is_valid_state(neighbor_state):
                        continue
                    neighbor = clip_state(neighbor_state)
                    n_trig = execute_Tr(*neighbor)
                    if not n_trig:
                        continue
                    rob += jaccard_similarity(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"prioritized_path{path_id}_isolated.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Prioritized Isolated Path {path_id}\n")
            f.write("dx dy dz\tScore\tSimilarity\tLengthDiff\tRobustness\tQ_value_complement\n")
            for s in samples:
                dx, dy, dz = s[0]
                f.write(f"{dx} {dy} {dz}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\t{s[5]:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_prioritized"
    for path_idx in isolated_group_indices:
        path = targetPaths[path_idx]
        samples = []
        attempts = 0
        while len(samples) < top_k and attempts < num_total * 5:
            attempts += 1
            state = (
                np.random.randint(dx_min, dx_max + 1),
                np.random.randint(dy_min, dy_max + 1),
                np.random.randint(dz_min, dz_max + 1)
            )
            triggered = execute_Tr(*state)
            if not triggered:
                continue
            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            q_value_complement = compute_q_value_normalized_complement(state, agent_similar)
            score = 0.28 * sim + 0.1 * len_diff + 0.19 * rob + 0.43 * q_value_complement
            samples.append((state, score, sim, len_diff, rob, q_value_complement))
        if samples:
            samples.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples[:top_k], base_dir=base_dir)


# === 孤立路径训练 ===
def prioritized_generate_and_train_for_isolated_paths(agent_similar, agent_isolated, similar_group,
                                                      isolated_group, path_documents, episodes=500, batch_size=32,
                                                      is_isolated=True):
    trained_paths = set()
    global_replay_count = 0

    stage1_samples_pool = {}
    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_similar.replay_buffer.get_high_reward_samples(target_path, num_samples=100)
        stage1_samples_pool[path_idx] = high_reward_samples

    for episode in range(episodes):
        episode_reward = 0
        episode_similarities = []
        episode_td_errors = []

        for path_idx in isolated_group:
            if path_idx in trained_paths:
                continue

            file_path = os.path.join(path_documents, f"prioritized_path{path_idx + 1}_isolated.txt")
            stage2_path_data = load_path_data(file_path)
            stage1_samples = stage1_samples_pool.get(path_idx, [])
            target_path = targetPaths[path_idx]

            # 训练参数设置
            BATCH_SIZE = 50
            N_SAMPLES_STAGE2 = min(140, len(stage2_path_data))
            N_SAMPLES_STAGE1 = min(60, len(stage1_samples))
            N_STEPS = 3
            N_BATCHES_STAGE2 = (N_SAMPLES_STAGE2 + BATCH_SIZE - 1) // BATCH_SIZE
            N_BATCHES_STAGE1 = (N_SAMPLES_STAGE1 + BATCH_SIZE - 1) // BATCH_SIZE
            PATH_REPEAT = 5

            for repeat in range(PATH_REPEAT):
                # 阶段2：使用新生成的孤立路径样本
                for batch_idx in range(N_BATCHES_STAGE2):
                    batch_start = batch_idx * BATCH_SIZE
                    batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES_STAGE2)

                    for test_data_idx in range(batch_start, batch_end):
                        if test_data_idx >= len(stage2_path_data):
                            break

                        state = stage2_path_data[test_data_idx]
                        prev_state = None
                        prev_triggered = None
                        prev_reward = None

                        for step in range(N_STEPS):
                            legal_actions = []
                            for a in range(agent_isolated.action_dim):
                                ddx, ddy, ddz = agent_isolated.decode_action(a)
                                cand_next = (state[0] + ddx, state[1] + ddy, state[2] + ddz)
                                if is_valid_state(cand_next):
                                    legal_actions.append(a)

                            if not legal_actions:
                                break

                            if random.random() < agent_isolated.epsilon:
                                action = random.choice(legal_actions)
                            else:
                                # 贪婪动作选择
                                normalized_state = normalize_state(state)
                                state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(
                                    device)
                                with torch.no_grad():
                                    q_values = agent_isolated.model(state_tensor)[0]
                                action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                            ddx, ddy, ddz = agent_isolated.decode_action(action)
                            next_state = clip_state((state[0] + ddx, state[1] + ddy, state[2] + ddz))

                            triggered = execute_Tr(*next_state)
                            reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)

                            if target_path.issubset(triggered):
                                reward += 2.0

                            done = (step == N_STEPS - 1)

                            td_error = agent_isolated.store_transition(state, action, reward, next_state, done)

                            priority_stats = agent_isolated.replay_buffer.get_priority_statistics()
                            current_priority = priority_stats['mean_priority'] if priority_stats else 0

                            prioritized_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                    current_priority, None)

                            episode_similarities.append(jaccard_similarity(triggered, target_path))
                            episode_td_errors.append(td_error)

                            if prev_reward is not None:
                                prioritized_metrics.record_action_improvement(reward, prev_reward)

                            prev_state = state
                            prev_triggered = triggered
                            prev_reward = reward
                            state = next_state
                            episode_reward += reward

                    if len(agent_isolated.replay_buffer) >= batch_size:
                        agent_isolated.train(batch_size)
                        global_replay_count += 1

                        if global_replay_count % 2 == 0:
                            agent_isolated.update_target_model()

                # 阶段1：使用相似路径智能体的高奖励样本
                if stage1_samples:
                    for batch_idx in range(N_BATCHES_STAGE1):
                        batch_start = batch_idx * BATCH_SIZE
                        batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES_STAGE1)

                        for sample_idx in range(batch_start, batch_end):
                            if sample_idx >= len(stage1_samples):
                                break

                            stage1_state_tuple, _, _, _ = stage1_samples[sample_idx]
                            state = stage1_state_tuple
                            prev_state = None
                            prev_triggered = None
                            prev_reward = None

                            for step in range(N_STEPS):
                                legal_actions = []
                                for a in range(agent_isolated.action_dim):
                                    ddx, ddy, ddz = agent_isolated.decode_action(a)
                                    cand_next = (state[0] + ddx, state[1] + ddy, state[2] + ddz)
                                    if is_valid_state(cand_next):
                                        legal_actions.append(a)

                                if not legal_actions:
                                    break

                                if random.random() < agent_isolated.epsilon:
                                    action = random.choice(legal_actions)
                                else:
                                    # 贪婪动作选择
                                    normalized_state = normalize_state(state)
                                    state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(
                                        device)
                                    with torch.no_grad():
                                        q_values = agent_isolated.model(state_tensor)[0]
                                    action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                                ddx, ddy, ddz = agent_isolated.decode_action(action)
                                next_state = clip_state((state[0] + ddx, state[1] + ddy, state[2] + ddz))

                                triggered = execute_Tr(*next_state)
                                reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                                reward *= 0.8

                                done = (step == N_STEPS - 1)

                                td_error = agent_isolated.store_transition(state, action, reward, next_state, done)

                                priority_stats = agent_isolated.replay_buffer.get_priority_statistics()
                                current_priority = priority_stats['mean_priority'] if priority_stats else 0

                                prioritized_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                        current_priority, None)

                                episode_similarities.append(jaccard_similarity(triggered, target_path))
                                episode_td_errors.append(td_error)

                                if prev_reward is not None:
                                    prioritized_metrics.record_action_improvement(reward, prev_reward)

                                prev_state = state
                                prev_triggered = triggered
                                prev_reward = reward
                                state = next_state
                                episode_reward += reward

                        if len(agent_isolated.replay_buffer) >= batch_size:
                            agent_isolated.train(batch_size)
                            global_replay_count += 1

                            if global_replay_count % 2 == 0:
                                agent_isolated.update_target_model()

            trained_paths.add(path_idx)

        avg_similarity = np.mean(episode_similarities) if episode_similarities else 0
        avg_td_error = np.mean(episode_td_errors) if episode_td_errors else 0
        priority_stats = agent_isolated.replay_buffer.get_priority_statistics()

        prioritized_metrics.record_episode_metrics(episode, episode_reward, avg_similarity, avg_td_error,
                                                   agent_isolated.epsilon, "isolated", priority_stats)

        if len(trained_paths) == len(isolated_group):
            break

    return agent_isolated


# === Excel保存函数 ===
def append_performance_metrics_to_excel(metrics_collector, filepath, run_number):
    """将性能指标追加到Excel文件"""
    avg_td_error = np.mean(metrics_collector.td_errors) if metrics_collector.td_errors else 0
    total_reward = metrics_collector.total_reward
    action_improvement_rate = np.mean(
        metrics_collector.action_improvements) if metrics_collector.action_improvements else 0
    avg_final_similarity = np.mean(
        metrics_collector.final_output_similarities) if metrics_collector.final_output_similarities else 0
    avg_episode_reward = np.mean(metrics_collector.episode_rewards) if metrics_collector.episode_rewards else 0
    reward_std = np.std(metrics_collector.episode_rewards) if metrics_collector.episode_rewards else 0
    training_time = metrics_collector.end_time - metrics_collector.start_time if metrics_collector.end_time and metrics_collector.start_time else 0
    avg_memory_usage = np.mean(metrics_collector.episode_memory_usage) if metrics_collector.episode_memory_usage else 0
    per_step_time = training_time / metrics_collector.step_count * 1000 if metrics_collector.step_count > 0 else 0
    avg_priority = np.mean(metrics_collector.priority_statistics) if metrics_collector.priority_statistics else 0
    avg_importance_weight = np.mean(metrics_collector.importance_weights) if metrics_collector.importance_weights else 0

    new_row = {
        'Run': f"Run {run_number}",
        '最终样本平均相似度': f"{avg_final_similarity:.4f}",
        '总奖励': f"{total_reward:,.2f}",
        '平均回合奖励': f"{avg_episode_reward:,.4f}",
        '奖励标准差': f"{reward_std:,.4f}",
        '平均TD误差': f"{avg_td_error:.4f}",
        '动作改进率': f"{action_improvement_rate:.4f}",
        '动作改进率(%)': f"{action_improvement_rate * 100:.2f}%",
        '总步数': f"{metrics_collector.step_count:,}",
        '最终样本数': f"{len(metrics_collector.final_output_similarities)}",
        '训练总时间(秒)': f"{training_time:.2f}",
        '训练总时间(分钟)': f"{training_time / 60:.2f}",
        '平均内存使用(MB)': f"{avg_memory_usage:.2f}",
        '每步时间(ms)': f"{per_step_time:.2f}",
        '平均优先级': f"{avg_priority:.4f}",
        '平均重要性权重': f"{avg_importance_weight:.4f}"
    }

    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    if os.path.exists(filepath):
        df = pd.read_excel(filepath, sheet_name='性能指标')
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    else:
        df = pd.DataFrame([new_row])

    with pd.ExcelWriter(filepath, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, sheet_name='性能指标', index=False)
        workbook = writer.book
        worksheet = writer.sheets['性能指标']

        worksheet.column_dimensions['A'].width = 15
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
            worksheet.column_dimensions[col].width = 18

        header_font = Font(bold=True, size=11)
        for cell in worksheet[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    print(f"运行 {run_number} 的性能指标已保存到: {filepath}")


def append_final_samples_to_excel(agent_similar, agent_isolated, similar_group, isolated_group, targetPaths, filepath,
                                  run_number):
    """将最终样本追加到Excel文件"""
    new_samples = []

    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_similar.replay_buffer.get_high_reward_samples(target_path, num_samples=20)

        for state_tuple, reward, sim, triggered in high_reward_samples:
            new_samples.append({
                'Run': f"Run {run_number}",
                '路径分组': '相似路径组',
                '路径ID': path_idx + 1,
                'dx': state_tuple[0],
                'dy': state_tuple[1],
                'dz': state_tuple[2],
                '相似度': f"{sim:.4f}",
                '奖励': f"{reward:.2f}",
                '触发分支数': len(triggered),
                '目标分支数': len(target_path),
                '触发分支': str(sorted(triggered)),
                '目标分支': str(sorted(target_path))
            })

    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)

        for state_tuple, reward, sim, triggered in high_reward_samples:
            new_samples.append({
                'Run': f"Run {run_number}",
                '路径分组': '孤立路径组',
                '路径ID': path_idx + 1,
                'dx': state_tuple[0],
                'dy': state_tuple[1],
                'dz': state_tuple[2],
                '相似度': f"{sim:.4f}",
                '奖励': f"{reward:.2f}",
                '触发分支数': len(triggered),
                '目标分支数': len(target_path),
                '触发分支': str(sorted(triggered)),
                '目标分支': str(sorted(target_path))
            })

    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    if os.path.exists(filepath):
        df = pd.read_excel(filepath, sheet_name='最终样本')
        df = pd.concat([df, pd.DataFrame(new_samples)], ignore_index=True)
    else:
        df = pd.DataFrame(new_samples)

    with pd.ExcelWriter(filepath, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, sheet_name='最终样本', index=False)
        workbook = writer.book
        worksheet = writer.sheets['最终样本']

        column_widths = {
            'A': 12, 'B': 15, 'C': 12, 'D': 10, 'E': 10, 'F': 10,
            'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 40, 'L': 40
        }
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        header_font = Font(bold=True, size=11)
        for cell in worksheet[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    print(f"运行 {run_number} 的最终样本已保存到: {filepath}")


# === 单次实验运行 ===
def run_single_experiment(run_number, results_save_dir):
    """运行单次实验"""
    print(f"\n{'=' * 80}")
    print(f"开始运行第 {run_number} 次实验（优先级DQN）")
    print(f"{'=' * 80}\n")

    # 重置指标收集器
    prioritized_metrics.reset()
    prioritized_metrics.start_training()

    model_path_similar = os.path.join(results_save_dir, f"run{run_number}_similar.pth")
    model_path_isolated = os.path.join(results_save_dir, f"run{run_number}_isolated.pth")
    path_documents = r"D:\Experiment\CNN\DQNNEW\path_samples_prioritized"

    similar_group, isolated_group = group_paths_by_similarity(targetPaths)

    # 确保样本目录存在
    os.makedirs(path_documents, exist_ok=True)

    # 检查相似路径样本文件是否已存在
    samples_exist = True
    for path_idx in similar_group:
        file_path = os.path.join(path_documents, f"prioritized_path{path_idx + 1}.txt")
        if not os.path.exists(file_path):
            samples_exist = False
            break

    if not samples_exist:
        print("生成相似路径样本...")
        generate_samples_for_similar_paths(similar_group, num_total=2000, top_k=200)
    else:
        print("相似路径样本文件已存在，跳过生成...")

    replay_buffer = PrioritizedExperienceReplay(capacity=10000, alpha=0.6, beta_start=0.4, beta_frames=100000)
    state_dim = 3
    action_dim = 30
    agent = PrioritizedDQNAgent(state_dim, action_dim, replay_buffer)

    agent = prioritized_generate_and_train_for_similar_paths(agent, similar_group, path_documents, episodes=500,
                                                             batch_size=32, is_isolated=False)

    os.makedirs(os.path.dirname(model_path_similar), exist_ok=True)
    torch.save({
        'model_state_dict': agent.model.state_dict(),
        'optimizer_state_dict': agent.optimizer.state_dict(),
        'epsilon': agent.epsilon
    }, model_path_similar)

    # 检查孤立路径样本文件是否已存在
    isolated_samples_exist = True
    for path_idx in isolated_group:
        file_path = os.path.join(path_documents, f"prioritized_path{path_idx + 1}_isolated.txt")
        if not os.path.exists(file_path):
            isolated_samples_exist = False
            break

    if not isolated_samples_exist:
        print("生成孤立路径样本...")
        generate_samples_for_isolated_paths_prioritized(agent, isolated_group, num_total=2000, top_k=200)
    else:
        print("孤立路径样本文件已存在，跳过生成...")

    # 初始化孤立路径智能体
    isolated_replay_buffer = PrioritizedExperienceReplay(capacity=15000, alpha=0.6, beta_start=0.4, beta_frames=100000)
    agent_isolated = PrioritizedDQNAgent(state_dim, action_dim, isolated_replay_buffer)

    try:
        checkpoint = torch.load(model_path_similar)
        agent_isolated.model.load_state_dict(checkpoint['model_state_dict'])
        agent_isolated.target_model.load_state_dict(checkpoint['model_state_dict'])
        agent_isolated.epsilon = checkpoint.get('epsilon', 0.5)
    except Exception as e:
        pass

    agent_isolated = prioritized_generate_and_train_for_isolated_paths(
        agent_similar=agent,
        agent_isolated=agent_isolated,
        similar_group=similar_group,
        isolated_group=isolated_group,
        path_documents=path_documents,
        episodes=500,
        batch_size=32,
        is_isolated=True
    )

    os.makedirs(os.path.dirname(model_path_isolated), exist_ok=True)
    torch.save({
        'model_state_dict': agent_isolated.model.state_dict(),
        'optimizer_state_dict': agent_isolated.optimizer.state_dict(),
        'epsilon': agent_isolated.epsilon
    }, model_path_isolated)

    # 结束训练计时
    prioritized_metrics.end_training()

    # 记录最终样本
    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, _, sim, triggered in high_reward_samples:
            prioritized_metrics.record_final_output_sample(triggered, target_path)

    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, _, sim, triggered in high_reward_samples:
            prioritized_metrics.record_final_output_sample(triggered, target_path)

    # 保存到Excel
    performance_excel_path = os.path.join(results_save_dir, "性能指标.xlsx")
    samples_excel_path = os.path.join(results_save_dir, "最终样本.xlsx")

    append_performance_metrics_to_excel(prioritized_metrics, performance_excel_path, run_number)
    append_final_samples_to_excel(agent, agent_isolated, similar_group, isolated_group, targetPaths, samples_excel_path,
                                  run_number)

    # 输出本次运行指标
    avg_similarity = np.mean(
        prioritized_metrics.final_output_similarities) if prioritized_metrics.final_output_similarities else 0
    training_time = prioritized_metrics.end_time - prioritized_metrics.start_time
    avg_priority = np.mean(prioritized_metrics.priority_statistics) if prioritized_metrics.priority_statistics else 0
    print(f"\n第 {run_number} 次运行完成:")
    print(f"  平均相似度: {avg_similarity:.4f}")
    print(f"  训练时间: {training_time:.2f} 秒")
    print(f"  总步数: {prioritized_metrics.step_count}")
    print(f"  平均优先级: {avg_priority:.4f}")


if __name__ == "__main__":
    results_save_dir = r"D:\Experiment\CNN\DQNNEW\results\prioritized_results"
    os.makedirs(results_save_dir, exist_ok=True)

    # 设置运行次数
    NUM_RUNS = 20

    print("=" * 80)
    print(f"开始 {NUM_RUNS} 次优先级DQN实验")
    print("=" * 80)

    for run in range(1, NUM_RUNS + 1):
        try:
            run_single_experiment(run, results_save_dir)
        except Exception as e:
            print(f"\n第 {run} 次运行出错: {str(e)}")
            import traceback

            traceback.print_exc()
            continue

    print("\n" + "=" * 80)
    print(f"所有 {NUM_RUNS} 次运行完成")
    print(f"结果保存在: {results_save_dir}")
    print("=" * 80)