import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """安全除法"""
    if denominator == 0:
        return default
    return numerator / denominator


def execute_Tr(x, y, z):
    """执行105分支逻辑，返回触发的分支集合"""
    # 初始化分支覆盖数组
    b = [0] * 105  # 根据分支数量调整大小

    # --- 分支 1-14 ---
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 1) > 120): b[0] = 1
    if ((y * z) / (x + 1) > 100) != ((y * y) / (x + 1) > 100): b[1] = 2
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 8) > 100): b[2] = 3
    if ((y * z) / (x + 1) > 100) != ((y + z) / (x + 1) > 100): b[3] = 4
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x * 3 + 1) > 100): b[4] = 5
    if ((y * z) / (x + 1) > 100) != ((z * z) / (x + 1) > 100): b[5] = 6
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 1) > 80): b[6] = 7
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + y) > 100): b[7] = 8
    if ((y * z) / (x + 1) > 100) != ((y * 10) / (x + 1) > 100): b[8] = 9
    if ((y * z) / (x + 1) > 100) != ((x * z) / (x + 1) > 100): b[9] = 10
    if ((y * z) / (x + 1) > 100) != ((y * z * 2) / (x + 1) > 100): b[10] = 11
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 5) > 100): b[11] = 12
    if ((y * z) / (x + 1) > 100) != ((y * z) / (x + 1) > 200): b[12] = 13
    if ((y * z) / (x + 1) > 100) != ((5 * z) / (x + 1) > 100): b[13] = 14

    # --- 分支 15-22 ---
    if ((z - x) < 0.3 * y) != ((z * 1.2 - x) < 0.3 * y): b[14] = 15
    if ((z - x) < 0.3 * y) != ((z - x * 1.1) < 0.3 * y): b[15] = 16
    if ((z - x) < 0.3 * y) != ((z - x) < 0.5 * y): b[16] = 17
    if ((z - x) < 0.3 * y) != ((z - x) < 0.3 * z): b[17] = 18
    if ((z - x) < 0.3 * y) != ((z - x) < 0.3 * y * y): b[18] = 19
    if ((z - x) < 0.3 * y) != ((z - y) < 0.3 * y): b[19] = 20
    if ((z - x) < 0.3 * y) != ((z - x) < 0.8 * y): b[20] = 21
    if ((z - x) < 0.3 * y) != ((z - x) < 0.3 * x): b[21] = 22

    # --- 分支 23-36 ---
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 3): b[22] = 23
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 * y ** 3) < z ** 4): b[23] = 24
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2 + y ** 3) < z ** 2): b[24] = 25
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 2) < z ** 2): b[25] = 26
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * 3) < (z ** 2)): b[26] = 27
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * x ** 3) < z ** 2): b[27] = 28
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 1 + y ** 3) < z ** 2): b[28] = 29
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 1) < z ** 2): b[29] = 30
    if ((x ** 3 + y ** 3) < z ** 2) != ((x * 3 + y ** 3) < z ** 2): b[30] = 31
    if ((x ** 3 + y ** 3) < z ** 2) != ((x * 3 + y ** 3) < z ** 2): b[31] = 32
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * 3) < z ** 2): b[32] = 33
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** x + y ** 3) < z ** 2): b[33] = 34
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2 + y ** 3) < z ** 2): b[34] = 35
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 1.5): b[35] = 36

    # --- 分支 37-54 ---
    cond_ratio = ((x / (y + 0.001)) > 5 and (y / (z + 0.001)) < 0.2)
    if cond_ratio != ((x / (y + 0.001)) > 5 or (y / (z + 0.001) < 0.2)): b[36] = 37
    if cond_ratio != ((x / (y + 0.001)) > 5 and (y % (z + 0.001) < 0.2)): b[37] = 38
    if cond_ratio != ((x / (y + 0.001)) > 5 and (y / (z - 0.001) < 0.4)): b[38] = 39
    if cond_ratio != ((x / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.3)): b[39] = 40
    if cond_ratio != ((x / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.1)): b[40] = 41
    if cond_ratio != ((x / (y + 0.001)) > 8 and (y / (z + 0.001) < 0.2)): b[41] = 42
    if cond_ratio != ((x / (y + 0.001)) > 2 and (y / (z + 0.001) < 0.2)): b[42] = 43
    if cond_ratio != ((x / (y + 0.001)) > 1 and (y / (z + 0.001) < 0.2)): b[43] = 44
    if cond_ratio != ((x / (y + 0.001)) > 10 and (y / (z + 0.001) < 0.2)): b[44] = 45
    if cond_ratio != ((x * y / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.2)): b[45] = 46
    if cond_ratio != ((x * x / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.2)): b[46] = 47
    if cond_ratio != ((x / (y + 0.001)) > 5 and (y * x / (z + 0.001) < 0.2)): b[47] = 48
    if cond_ratio != ((x / (y + 0.001)) > 5 and (y * y / (z + 0.001) < 0.2)): b[48] = 49
    if cond_ratio != ((x / (y + 0.001)) > 5 and (y * z / (z + 0.001) < 0.2)): b[49] = 50
    if cond_ratio != ((x / (y + 0.001)) > 5 and (y / (z + 0.001) < 0.5)): b[50] = 51
    if cond_ratio != ((x / (y + 0.001)) > 5 and (y - (z + 0.001) < 0.2)): b[51] = 52
    if cond_ratio != ((x + (y + 0.001)) > 5 and (y / (z + 0.001) < 0.2)): b[52] = 53
    if cond_ratio != ((x - (y + 0.001)) > 5 and (y / (z + 0.001) < 0.2)): b[53] = 54

    # --- 分支 55-75 ---
    cond_abs = (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 5)
    if cond_abs != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 20): b[54] = 55
    if cond_abs != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 12.9): b[55] = 56
    if cond_abs != (abs(x - y) > 10 and abs(y - z) > 10 and abs(20 - z) < 5): b[56] = 57
    if cond_abs != (abs(x - y) > 10 and abs(y - 80) > 10 and abs(x - z) < 5): b[57] = 58
    if cond_abs != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x + z) < 5): b[58] = 59
    if cond_abs != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 7): b[59] = 60
    if cond_abs != (abs(x - y) > 10 and abs(y - z) > 15 and abs(10 - z) < 10): b[60] = 61
    if cond_abs != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - 8) < 5): b[61] = 62
    if cond_abs != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - 2) < 5): b[62] = 63
    if cond_abs != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x * 4 - z) < 5): b[63] = 64
    if cond_abs != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x * z) < 5): b[64] = 65
    if cond_abs != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x + z) < 5): b[65] = 66
    if cond_abs != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x * y - z) < 5): b[66] = 67
    if cond_abs != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 2): b[67] = 68
    if cond_abs != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - y) < 5): b[68] = 69
    if cond_abs != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z) < 15): b[69] = 70
    if cond_abs != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x * 2 - z) < 5): b[70] = 71
    if cond_abs != (abs(x - y) > 10 and abs(y - z) > 10 and abs(y - z) < 5): b[71] = 72
    if cond_abs != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x * z - z) < 5): b[72] = 73
    if cond_abs != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x - z * 2) < 5): b[73] = 74
    if cond_abs != (abs(x - y) > 10 and abs(y - z) > 10 and abs(x * x - z) < 5): b[74] = 75

    # --- 分支 76-85 ---
    cond_extreme = ((x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2))
    if cond_extreme != ((x * x > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): b[75] = 76
    if cond_extreme != ((x * y > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): b[76] = 77
    if cond_extreme != ((x * z > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): b[77] = 78
    if cond_extreme != ((x > 90 or x < 5) and (y * y > 80 or y < 3) and (z > 75 or z < 2)): b[78] = 79
    if cond_extreme != ((x > 90 or x < 5) and (y * z > 80 or y < 3) and (z > 75 or z < 2)): b[79] = 80
    if cond_extreme != ((x > 90 or x < 5) and (y * x > 80 or y < 3) and (z > 75 or z < 2)): b[80] = 81
    if cond_extreme != ((x * 10 > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): b[81] = 82
    if cond_extreme != ((x * 15 > 90 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): b[82] = 83
    if cond_extreme != ((x > 50 or x < 5) and (y > 80 or y < 3) and (z > 75 or z < 2)): b[83] = 84
    if cond_extreme != ((x > 90 or x < 5) and (y > 80 or y < 3) and (z * 40 > 75 or z < 2)): b[84] = 85

    # --- 分支 86-100 ---
    cond_mixed = ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2)
    if cond_mixed != ((x * x) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): b[85] = 86
    if cond_mixed != ((x * 60) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): b[86] = 87
    if cond_mixed != ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): b[87] = 88
    if cond_mixed != ((70 * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): b[88] = 89
    if cond_mixed != ((x * y) + (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): b[89] = 90
    if cond_mixed != ((x * y) / (z + 1) > 70 and x ** 2 + y ** 2 > z ** 2): b[90] = 91
    if cond_mixed != ((x * y) / (z + 1) > 80 and x ** 2 + y ** 2 > z ** 2): b[91] = 92
    if cond_mixed != ((x * y) / (z + 1) > 50 and x ** 1.5 + y ** 2 > z ** 2): b[92] = 93
    if cond_mixed != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2.2): b[93] = 94
    if cond_mixed != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2.5): b[94] = 95
    if cond_mixed != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 1.5 > z ** 2): b[95] = 96
    if cond_mixed != ((x * y) / (z + 1) > 80 and x ** 2 + y ** 2 > z ** 2): b[96] = 97
    if cond_mixed != ((x * y) / (z + 1) > 65 and x ** 2 + y ** 2 > z ** 2): b[97] = 98
    if cond_mixed != ((x * y) / (z + 1) > 50 and x ** 1.2 + y ** 2 > z ** 2): b[98] = 99
    if cond_mixed != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 1.2 > z ** 2): b[99] = 100

    # --- 分支 101-105 ---
    cond_sqrt = (z ** 0.5 > (x + y) / 2 and x * y * z > 1000)
    if cond_sqrt != (z ** 0.7 > (x + y) / 2 and x * y * z > 1000): b[100] = 101
    if cond_sqrt != (z ** 0.5 > (x + y) / 6 and x * y * z > 1000): b[101] = 102
    if cond_sqrt != (z ** 0.8 > (x + y) / 2 and x * y * z > 1000): b[102] = 103
    if cond_sqrt != (z ** 0.5 > (x + x) / 2 and x * y * z > 1000): b[103] = 104
    if cond_sqrt != (z ** 0.5 > (y + y) / 2 and x * y * z > 1000): b[104] = 105

    triggered = set()
    for i, val in enumerate(b):
        if val > 0:
            triggered.add(i + 1)  # 分支编号从1开始
    return triggered


def execute_validation_rules(dx: int, dy: int, dz: int) -> Set[int]:
    """执行验证规则，返回触发的分支集合（调用105分支版本）"""
    return execute_Tr(dx, dy, dz)


def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """计算适应度"""
    generated_path = execute_validation_rules(int(particle[0]), int(particle[1]), int(particle[2]))

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """基本PSO优化器"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # 范围: x(2,100), y(1,105), z(1,110)
        self.bounds = bounds if bounds else [(2, 100), (1, 105), (1, 110)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """初始化粒子"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """更新速度和位置"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """对目标路径进行PSO优化"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(
                            int(particles[i][0]), int(particles[i][1]), int(particles[i][2])
                        ),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(
                int(gbest_particle[0]), int(gbest_particle[1]), int(gbest_particle[2])
            ),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """对所有路径运行PSO"""

    print(f"\n{'=' * 70}")
    print(f"Baseline PSO - 路径优化")
    print(f"{'=' * 70}")
    print(f"粒子数: {n_particles}, 最大迭代次数: {max_iterations}")
    print(f"路径数量: {len(target_paths)}")
    print(f"范围: x(2,100), y(1,105), z(1,110)")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"路径 {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "完美求解" if result['success'] else f"(适应度: {result['best_fitness']:.3f})"
        print(f"{status} | 耗时: {result['time']:.2f}s | 迭代次数: {result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f"求解成功: {success_count}/{len(target_paths)} ({success_rate:.1f}%) | 总耗时: {total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """多次运行实验"""

    print(f"\n{'=' * 70}")
    print(f"Baseline PSO - {num_runs}次运行")
    print(f"{'=' * 70}")
    print(f"粒子数: {n_particles}, 最大迭代次数: {max_iterations}, 路径数: {len(target_paths)}")
    print(f"范围: x(2,100), y(1,105), z(1,110)")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"--- 第 {run_idx}/{num_runs} 次运行 ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f"求解成功: {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs}次运行完成 | 总耗时: {total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """导出Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # Sheet 1: 运行汇总
    ws1 = wb.active
    ws1.title = "运行汇总"
    ws1.sheet_view.showGridLines = False

    headers = ["运行", "成功率", "求解/总数", "平均适应度", "平均迭代次数", "耗时(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f"运行 {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # Sheet 2: 路径统计
    ws2 = wb.create_sheet(title="路径统计")
    ws2.sheet_view.showGridLines = False

    headers2 = ["路径ID", "求解/总数", "成功率", "平均适应度", "平均迭代次数", "最小迭代次数", "最大迭代次数"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"路径 {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # Sheet 3: 详细结果
    ws3 = wb.create_sheet(title="详细结果")
    ws3.sheet_view.showGridLines = False

    headers3 = ["路径ID", "运行", "位置(x,y,z)", "适应度", "迭代次数", "触发路径"]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"Run {run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # Sheet 4: 目标路径
    ws4 = wb.create_sheet(title="目标路径")
    ws4.sheet_view.showGridLines = False

    headers4 = ["路径ID", "目标路径", "分支数"]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"路径 {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f"Excel已保存: {filename}")
    print(f"{'=' * 70}")
    print(f"包含工作表:")
    print(f"  1. 运行汇总      - {len(all_results)}次运行汇总")
    print(f"  2. 路径统计      - 每条路径的统计")
    print(f"  3. 详细结果      - 每次运行每条路径的详细结果")
    print(f"  4. 目标路径      - 目标路径定义")
    print(f"{'=' * 70}\n")

    return filename


def main():
    # 目标路径定义（105分支版本）
    target_paths = [
        {15, 20, 29, 31, 32, 38, 41, 42, 45, 48, 49, 50, 57, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 79,
         80, 81, 86, 87, 90},  # A1
        {6, 10, 15, 20, 23, 29, 31, 32, 38, 41, 42, 45, 48, 49, 50, 57, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73,
         74, 75, 79, 80, 81},  # A2
        {1, 3, 4, 5, 6, 8, 9, 10, 13, 14, 57, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 76, 77, 78, 82,
         83, 84, 94, 95, 100},  # A3
        {16, 17, 18, 19, 21, 22, 23, 29, 31, 32, 38, 41, 45, 48, 49, 50, 57, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72,
         73, 74, 75, 105},  # A4
        {6, 7, 10, 11, 15, 57, 58, 59, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 79, 80, 81, 92, 93, 94, 95,
         96, 97, 99, 100},  # A5
        {1, 2, 3, 4, 5, 8, 9, 10, 13, 14, 16, 17, 18, 19, 20, 21, 23, 76, 77, 78, 82, 83, 84, 87, 91, 92, 93, 94, 95, 96,
         97, 98, 99, 100},  # A6
        {11, 15, 20, 57, 59, 61, 62, 63, 64, 65, 66, 67, 69, 71, 72, 73, 74, 75, 79, 80, 81, 88, 89, 91, 92, 93, 94, 95, 97,
         98, 99, 100},  # A7
        {1, 3, 4, 5, 6, 8, 9, 10, 12, 13, 14, 18, 22, 57, 58, 59, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 86, 87,
         91, 92, 97, 98},  # A8
        {16, 18, 19, 21, 22, 23, 24, 25, 29, 31, 32, 35, 38, 48, 50, 57, 59, 61, 62, 63, 64, 65, 66, 67, 69, 71, 72, 73, 74,
         75, 105},  # A9
        {15, 20, 29, 31, 32, 38, 41, 42, 45, 48, 49, 50, 59, 62, 63, 64, 65, 66, 67, 68, 69, 71, 72, 73, 74, 75, 87, 90,
         102, 105},  # A10
        {1, 2, 3, 4, 5, 8, 9, 10, 12, 13, 14, 19, 23, 24, 25, 26, 27, 29, 30, 31, 32, 33, 35, 37, 46, 47, 53, 101, 102,
         103},  # A11
        {3, 4, 5, 6, 8, 10, 12, 13, 14, 18, 22, 57, 59, 62, 64, 65, 66, 67, 68, 69, 72, 73, 74, 85, 86, 87, 91, 92, 97, 98},
        # A12
        {15, 20, 23, 29, 31, 32, 37, 43, 44, 46, 47, 53, 54, 57, 59, 61, 62, 63, 64, 65, 66, 67, 69, 71, 72, 73, 74, 75},
        # A13
        {6, 10, 16, 18, 19, 21, 22, 23, 25, 29, 31, 32, 35, 38, 45, 48, 49, 50, 55, 56, 60, 70, 79, 80, 81, 105},  # A14
        {3, 4, 5, 6, 8, 9, 10, 12, 13, 14, 17, 19, 20, 21, 27, 30, 33, 55, 57, 61, 63, 64, 70, 88, 89, 90, 104},  # A15
        {2, 3, 4, 5, 8, 9, 10, 14, 19, 23, 24, 26, 27, 28, 30, 33, 62, 63, 67, 101, 102, 103, 104},  # A16
        {1, 2, 3, 4, 5, 8, 9, 10, 12, 13, 14, 19, 28, 34, 36, 37, 46, 47, 53, 101, 102, 103},  # A17
        {15, 20, 29, 31, 32, 37, 39, 40, 51, 52, 55, 56, 70, 79, 80, 81, 86, 87, 90}  # A18
    ]

    print("=" * 70)
    print("Baseline PSO - 105分支版本")
    print("=" * 70)
    print(f"范围: x(2,100), y(1,105), z(1,110)")
    print(f"路径数: {len(target_paths)}")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("程序完成")


if __name__ == "__main__":
    main()