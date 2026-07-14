import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === 全局取值范围配置 ===
MIN_X = 2
MAX_X = 100
MIN_Y = 1
MAX_Y = 150
MIN_Z = 1
MAX_Z = 200


def normalize_state(state):
    """将状态归一化到 [0, 1] 区间"""
    weather_norm = (state[0] - MIN_X) / (MAX_X - MIN_X)
    time_norm = (state[1] - MIN_Y) / (MAX_Y - MIN_Y)
    z_norm = (state[2] - MIN_Z) / (MAX_Z - MIN_Z)
    return (weather_norm, time_norm, z_norm)


def denormalize_state(state_norm):
    """将归一化状态还原"""
    weather = int(round(state_norm[0] * (MAX_X - MIN_X) + MIN_X))
    time_period = int(round(state_norm[1] * (MAX_Y - MIN_Y) + MIN_Y))
    z = int(round(state_norm[2] * (MAX_Z - MIN_Z) + MIN_Z))

    # 边界保护
    weather = np.clip(weather, MIN_X, MAX_X)
    time_period = np.clip(time_period, MIN_Y, MAX_Y)
    z = np.clip(z, MIN_Z, MAX_Z)

    return (weather, time_period, z)


def generate_input():
    return [
        random.randint(MIN_X, MAX_X),
        random.randint(MIN_Y, MAX_Y),
        random.randint(MIN_Z, MAX_Z)
    ]


def execute_Tr(x, y, z):
    triggered = set()

    # --- 分支 1-11 (原 energy_y * energy_z / (energy_x + 1) > 140 的变异) ---
    if ((y * z) / (x + 1) > 140) != ((y * y) / (x + 1) > 140): triggered.add(1)
    if ((y * z) / (x + 1) > 140) != ((z * z) / (x + 1) > 140): triggered.add(2)
    if ((y * z) / (x + 1) > 140) != ((y * x) / (x + 1) > 140): triggered.add(3)
    if ((y * z) / (x + 1) > 140) != ((y * z) / (x + 3) > 140): triggered.add(4)
    if ((y * z) / (x + 1) > 140) != ((y * z) / (x - 1) > 140): triggered.add(5)
    if ((y * z) / (x + 1) > 140) != ((y * z * 2) / (y + 1) > 140): triggered.add(6)
    if ((y * z) / (x + 1) > 140) != ((y * z) / (x + 1) > 100): triggered.add(7)
    if ((y * z) / (x + 1) > 140) != ((y * z) / (x + 1) > 180): triggered.add(8)
    if ((y * z) / (x + 1) > 140) != ((y * z) / (x + 10) > 140): triggered.add(9)
    if ((y * z) / (x + 1) > 140) != ((y * z) / (x * 1) > 140): triggered.add(10)
    if ((y * z) / (x + 1) > 140) != ((y * 30) / (x + 1) > 140): triggered.add(11)

    # --- 分支 12-21 (原 (energy_z - energy_x) < 0.22 * energy_y 的变异) ---
    if ((z - x) < 0.22 * y) != ((z - x) < 0.22 * x): triggered.add(12)
    if ((z - x) < 0.22 * y) != ((z - x) < 0.22 * z): triggered.add(13)
    if ((z - x) < 0.22 * y) != ((z - x) < 0.32 * y): triggered.add(14)
    if ((z - x) < 0.22 * y) != ((z - x) < 0.12 * y): triggered.add(15)
    if ((z - x) < 0.22 * y) != ((z * 2 - x) < 0.22 * y): triggered.add(16)
    if ((z - x) < 0.22 * y) != ((z - x * 1.2) < 0.22 * y): triggered.add(17)
    if ((z - x) < 0.22 * y) != ((z + x) < 0.22 * y): triggered.add(18)
    if ((z - x) < 0.22 * y) != ((z - 20) < 0.22 * y): triggered.add(19)
    if ((z - x) < 0.22 * y) != ((90 - x) < 0.22 * y): triggered.add(20)
    if ((z - x) < 0.22 * y) != ((z - x) < 0.4 * y): triggered.add(21)

    # --- 分支 22-32 (原 (energy_x^3 + energy_y^3) < energy_z^2 的变异) ---
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2.7 + y ** 3) < z ** 2): triggered.add(22)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 2.6) < z ** 2): triggered.add(23)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 1.8): triggered.add(24)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 - y ** 3) < z ** 2): triggered.add(25)
    if ((x ** 3 + y ** 3) < z ** 2) != ((y ** 3 + y ** 3) < z ** 2): triggered.add(26)
    if ((x ** 3 + y ** 3) < z ** 2) != ((z ** 3 + y ** 3) < z ** 2): triggered.add(27)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + x ** 3) < z ** 2): triggered.add(28)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + z ** 3) < z ** 2): triggered.add(29)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < x ** 2): triggered.add(30)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < y ** 2): triggered.add(31)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 2.5): triggered.add(32)

    # --- 分支 33-42 (原 x/(y+0.01)>5 and y/(z+0.01)<0.2 的变异) ---
    if ((x / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.2) != ((x / (z + 0.01)) > 5 and (y / (z + 0.01)) < 0.2): triggered.add(33)
    if ((x / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.2) != ((x / (x + 0.01)) > 5 and (y / (z + 0.01)) < 0.2): triggered.add(34)
    if ((x / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.2) != ((z / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.2): triggered.add(35)
    if ((x / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.2) != ((y / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.2): triggered.add(36)
    if ((x / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.2) != ((x / (y + 0.01)) > 5 and (z / (z + 0.01)) < 0.2): triggered.add(37)
    if ((x / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.2) != ((x / (y + 0.01)) > 5 and (x / (z + 0.01)) < 0.2): triggered.add(38)
    if ((x / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.2) != ((x / (y + 0.01)) > 5 and (y / (y + 0.01)) < 0.2): triggered.add(39)
    if ((x / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.2) != ((x / (y + 0.01)) > 5 and (y / (x + 0.01)) < 0.2): triggered.add(40)
    if ((x / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.2) != ((x / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.15): triggered.add(41)
    if ((x / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.2) != ((x / (y + 0.01)) > 7 and (y / (z + 0.01)) < 0.2): triggered.add(42)

    # --- 分支 43-52 (原 abs(x-y)>16 and abs(y-z)>18 and abs(x-z)<9 的变异) ---
    if (abs(x - y) > 16 and abs(y - z) > 18 and abs(x - z) < 9) != (abs(x * 1.2 - y) > 16 and abs(y - z) > 18 and abs(x - z) < 9): triggered.add(43)
    if (abs(x - y) > 16 and abs(y - z) > 18 and abs(x - z) < 9) != (abs(x - y) > 16 and abs(y * 2 - z) > 18 and abs(x - z) < 9): triggered.add(44)
    if (abs(x - y) > 16 and abs(y - z) > 18 and abs(x - z) < 9) != (abs(x - y) > 19 and abs(y - z) > 18 and abs(y - z) < 9): triggered.add(45)
    if (abs(x - y) > 16 and abs(y - z) > 18 and abs(x - z) < 9) != (abs(x - y) > 16 and abs(x - z) > 18 and abs(x - z) < 9): triggered.add(46)
    if (abs(x - y) > 16 and abs(y - z) > 18 and abs(x - z) < 9) != (abs(x - y) > 16 and abs(y - z) > 40 and abs(x - z) < 9): triggered.add(47)
    if (abs(x - y) > 16 and abs(y - z) > 18 and abs(x - z) < 9) != (abs(x - y) > 16 and abs(y - z) > 18 and abs(x * 2 - z) < 9): triggered.add(48)
    if (abs(x - y) > 16 and abs(y - z) > 18 and abs(x - z) < 9) != (abs(x - y) > 16 and abs(y - z * 0.2) > 18 and abs(x - z) < 9): triggered.add(49)
    if (abs(x - y) > 16 and abs(y - z) > 18 and abs(x - z) < 9) != (abs(x - y) > 16 and abs(y - z) > 18 and abs(x * 1.5 - z) < 9): triggered.add(50)
    if (abs(x - y) > 16 and abs(y - z) > 18 and abs(x - z) < 9) != (abs(x - y) > 16 and abs(y - z) > 18 and abs(x - z * 0.87) < 9): triggered.add(51)
    if (abs(x - y) > 16 and abs(y - z) > 18 and abs(x - z) < 9) != (abs(x - y) > 16 and abs(y - z) > 18 and abs(x - z) < 7.8): triggered.add(52)

    # --- 分支 53-63 (原 (x>95 or x<5) and (y>90 or y<3) and (z>85 or z<2) 的变异) ---
    if ((x > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)) != ((x > 95 or x < 5) and (y * y > 90 or y < 3) and (z > 85 or z < 2)): triggered.add(53)
    if ((x > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)) != ((x > 95 or x < 5) and (y * x > 90 or y < 3) and (z > 85 or z < 2)): triggered.add(54)
    if ((x > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)) != ((x > 95 or x < 5) and (y * z > 90 or y < 3) and (z > 85 or z < 2)): triggered.add(55)
    if ((x > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)) != ((x > 95 or x < 5) and (y * 80 > 90 or y < 3) and (z > 85 or z < 2)): triggered.add(56)
    if ((x > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)) != ((x * y > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)): triggered.add(57)
    if ((x > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)) != ((x * x > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)): triggered.add(58)
    if ((x > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)) != ((x * z > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)): triggered.add(59)
    if ((x > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)) != ((x * 50 > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)): triggered.add(60)
    if ((x > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)) != ((x > 95 or x < 5) and (y * 40 > 90 or y < 3) and (z > 85 or z < 2)): triggered.add(61)
    if ((x > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)) != ((x > 95 or x < 5) and (y > 90 or y < 3) and (z * z > 85 or z < 2)): triggered.add(62)
    if ((x > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)) != ((x * 40 > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)): triggered.add(63)

    # --- 分支 64-75 (原 x^0.7+y^0.7>z^0.9 and x+y+z<180 的变异) ---
    if (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180) != (x ** 0.6 + y ** 0.7 > z ** 0.9 and x + y + z < 180): triggered.add(64)
    if (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180) != (x ** 0.7 + y ** 0.7 > z ** 0.9 and z + y + z < 180): triggered.add(65)
    if (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180) != (x ** 0.7 + y ** 0.8 > z ** 0.9 and x + y + z < 180): triggered.add(66)
    if (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180) != (x ** 0.7 + y ** 0.7 > z ** 0.8 and x + y + z < 180): triggered.add(67)
    if (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180) != (x ** 0.7 + z ** 0.7 > z ** 0.9 and x + y + z < 180): triggered.add(68)
    if (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180) != (y ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180): triggered.add(69)
    if (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180) != (z ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180): triggered.add(70)
    if (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180) != (x ** 0.7 + x ** 0.7 > z ** 0.9 and x + y + z < 180): triggered.add(71)
    if (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180) != (x ** 0.7 + y ** 0.7 > x ** 0.9 and x + y + z < 180): triggered.add(72)
    if (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180) != (x ** 0.7 + y ** 0.7 > z ** 0.9 and y + y + z < 180): triggered.add(73)
    if (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180) != (x ** 0.7 + y ** 0.7 > z ** 0.9 and z + y + z < 180): triggered.add(74)
    if (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180) != (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + x + z < 180): triggered.add(75)

    # --- 分支 76-85 (原 (x+y)^1.3<z^1.6 and x+y+z/3>35 的变异) ---
    if ((x + y) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35) != ((y + y) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35): triggered.add(76)
    if ((x + y) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35) != ((z + y) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35): triggered.add(77)
    if ((x + y) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35) != ((x + x) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35): triggered.add(78)
    if ((x + y) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35) != ((x + z) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35): triggered.add(79)
    if ((x + y) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35) != ((x + 20) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35): triggered.add(80)
    if ((x + y) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35) != ((x + y) ** 1 < z ** 1.6 and x + y + z / 3 > 35): triggered.add(81)
    if ((x + y) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35) != ((x + y) ** 1.3 < z ** 1.7 and x + y + z / 3 > 35): triggered.add(82)
    if ((x + y) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35) != ((x + y) ** 1.2 < z ** 1.6 and x + y + z / 3 > 35): triggered.add(83)
    if ((x + y) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35) != ((x + y) ** 1.3 < z ** 1.6 and y + y + z / 3 > 35): triggered.add(84)
    if ((x + y) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35) != ((x + y) ** 1.3 < z ** 1.6 and x + y - z / 3 > 35): triggered.add(85)

    return triggered


# === target path definitions ===
targetPaths = [
    {2,3,6,8,9,11,12,13,15,16,18,19,20,25,48,50,65,68,71,73,74,78,79,80,81,82,83},  # A1
    {6,16,18,19,26,33,34,36,37,38,39,41,42,45,46,48,49,50,51,52,53,54,55,56,61},  # A2
    {2,3,4,6,8,9,11,12,13,15,16,18,20,25,48,50,68,71,73,78,79,80,81,82,83},  # A3
    {6,12,13,17,20,21,26,33,34,36,37,38,39,45,46,48,49,50,53,54,55,56,61},  # A4
    {1,16,18,19,20,25,45,46,47,48,50,64,65,68,71,73,74,78,79,80,81,82,83},  # A5
    {2,3,4,6,8,9,11,12,13,15,16,18,19,20,25,50,51,75,78,79,80,81,82,83},  # A6
    {12,13,17,20,26,33,34,36,37,38,39,45,46,48,49,50,51,57,58,59,60,63},  # A7
    {16,18,19,20,25,43,45,46,47,48,50,51,52,65,74,75,78,79,80,81,82,83},  # A8
    {16,18,19,20,25,43,45,46,47,48,50,64,68,71,73,77,78,79,80,81,82,83},  # A9
    {2,3,6,8,9,11,14,17,21,25,48,64,65,68,71,73,74,78,79,80,81,82,83},  # A10
    {1,7,12,13,16,18,19,20,25,50,51,65,68,71,73,74,78,79,80,81,82,83},  # A11
    {1,5,7,10,12,13,15,16,18,20,25,48,50,51,68,71,73,78,79,80,81},  # A12
    {12,13,17,20,26,33,34,36,37,38,39,45,46,48,50,58,59,60,63,84},  # A13
    {18,19,20,26,33,34,36,37,38,39,41,64,69,70,72,75,76,77,81,83},  # A14
    {16,18,19,20,26,33,34,35,36,37,38,39,41,67,68,71,78,79,80,84},  # A15
    {3,6,12,13,15,16,18,20,25,28,62,65,68,71,73,74,78,79,80,81},  # A16
    {16,18,19,20,26,33,34,36,37,38,39,67,68,71,78,79,80,84,85},  # A17
    {18,19,20,26,33,34,36,37,38,39,66,67,68,71,76,77,81,82,83},  # A18
    {2,6,24,27,28,29,30,31,33,34,36,37,39,58,59,60,63,85},  # A19
    {2,6,22,26,32,33,34,36,37,38,39,58,59,60,63,85},  # A20
    {12,13,14,17,21,26,44,45,46,47,48,49,50,84,85},  # A21
    {18,19,20,26,40,66,67,68,71,76,77,81,82,83},  # A22
    {3,23,25,28,32,35,53,54,55,56,61,85}  # A23
]

# 转换为集合列表
target_paths = [set(path) for path in targetPaths]

NUM_PATHS = len(targetPaths)


def jaccard_similarity(set1, set2):
    """Compute Jaccard similarity"""
    if not set2:
        return 0.0

    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0


# === reward function ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    """计算奖励"""
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10

    if target_path.issubset(triggered):
        reward += 1

    if prev_triggered is not None:
        prev_sim = jaccard_similarity(prev_triggered, target_path)
        improvement = sim - prev_sim
        reward += improvement * 5

    return reward


# === Path Similarity ===
def compute_path_similarity_matrix(paths):
    """Path Similarity"""
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix


def group_paths_by_similarity(paths, threshold_percentile=50):
    """基于相似度对路径进行分组"""
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.percentile(avg_sim_scores, threshold_percentile)

    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]

    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)

    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group


# === Sample generation ===
def compute_robustness(state, path, sample_size=9):
    """计算鲁棒性(修改版)"""
    base = execute_Tr(state[0], state[1], state[2])
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0

    # 邻域定义: 
    deltas = [
        (-1, -1, -6), (0, -1, 0), (1, -1, 6),
        (-1, 0, -6), (1, 0, 6),
        (-1, 1, -6), (0, 1, 0), (1, 1, 6),
        (0, 0, 0)
    ]

    for dw, dt, dz in deltas[:sample_size]:
        if dw == dt == dz == 0:
            continue

        neighbor_weather = int(np.clip(state[0] + dw, MIN_X, MAX_X))
        neighbor_time = int(np.clip(state[1] + dt, MIN_Y, MAX_Y))
        neighbor_z = int(np.clip(state[2] + dz, MIN_Z, MAX_Z))
        neighbor = (neighbor_weather, neighbor_time, neighbor_z)

        n_trig = execute_Tr(neighbor[0], neighbor[1], neighbor[2])
        if not n_trig:
            continue

        rob += jaccard_similarity(n_trig, base)
        neighbors += 1

    return rob / neighbors if neighbors > 0 else 0.0


def generate_samples_for_all_paths(num_candidates=2000, top_k=200, run_id=1):
    """为所有路径生成样本"""
    BEST_WEIGHTS = [0.55, 0.25, 0.2]

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_individual.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Individual Path {path_id} (Weighted Screening) - Run {run_id}\n")
            f.write("weather time_period z\tScore\tSimilarity\tRobustness\tLengthDiff\n")
            for s in samples:
                weather, time_period, z = s['state']
                f.write(
                    f"{weather} {time_period} {z}\t{s['score']:.4f}\t{s['similarity']:.4f}\t"
                    f"{s['robustness']:.4f}\t{s['length_diff']:.4f}\n"
                )

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_individual"

    for path_idx in range(len(targetPaths)):
        path = targetPaths[path_idx]
        candidate_samples = []
        attempts = 0
        max_attempts = num_candidates * 10

        while len(candidate_samples) < num_candidates and attempts < max_attempts:
            attempts += 1

            weather = np.random.randint(MIN_X, MAX_X + 1)
            time_period = np.random.randint(MIN_Y, MAX_Y + 1)
            z = np.random.randint(MIN_Z, MAX_Z + 1)
            state = (weather, time_period, z)

            triggered = execute_Tr(weather, time_period, z)
            if not triggered:
                continue

            sim = jaccard_similarity(triggered, path)
            rob = compute_robustness(state, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))

            candidate_samples.append({
                'state': state,
                'similarity': sim,
                'robustness': rob,
                'length_diff': len_diff,
                'triggered': triggered
            })

        if candidate_samples:
            for sample in candidate_samples:
                score = (BEST_WEIGHTS[0] * sample['similarity'] +
                         BEST_WEIGHTS[1] * sample['robustness'] +
                         BEST_WEIGHTS[2] * sample['length_diff'])
                sample['score'] = score

            candidate_samples.sort(key=lambda x: x['score'], reverse=True)
            selected_samples = candidate_samples[:top_k]
            save_samples(path_id=path_idx + 1, samples=selected_samples, base_dir=base_dir)


# === 共享经验回放池 ===
class SharedExperienceReplay:
    """共享经验回放池(带优先级)"""

    def __init__(self, capacity=20000):
        self.capacity = capacity
        self.buffer = deque(maxlen=self.capacity)
        self.priorities = deque(maxlen=self.capacity)

    def append(self, experience):
        """添加经验"""
        self.buffer.append(experience)
        self.priorities.append(experience[-1])

    def sample(self, batch_size, alpha=0.6):
        """优先级采样"""
        if len(self.buffer) < batch_size:
            return [], [], []

        priorities = np.array(self.priorities, dtype=np.float64)
        priorities = np.power(priorities, alpha)
        probabilities = priorities / np.sum(priorities)

        batch_indices = np.random.choice(len(self.buffer), batch_size, p=probabilities, replace=False)
        batch = [self.buffer[idx] for idx in batch_indices]

        return batch, batch_indices, probabilities[batch_indices]

    def update_priorities(self, batch_indices, td_errors):
        """更新优先级"""
        for idx, td_error in zip(batch_indices, td_errors):
            if idx < len(self.priorities):
                self.priorities[idx] = max(abs(td_error), 1e-6)

    def __len__(self):
        return len(self.buffer)

    def get_high_reward_samples(self, target_path, num_samples=20):
        """获取高奖励样本(去重, 重新计算)"""
        if len(self.buffer) == 0:
            return []

        samples_with_scores = []
        seen_states = set()  # 用于去重

        for experience in self.buffer:
            state_tensor = experience[0]
            state_norm = state_tensor.cpu().numpy().flatten()

            # 反归一化
            state_tuple = denormalize_state((state_norm[0], state_norm[1], state_norm[2]))

            # 去重
            if state_tuple in seen_states:
                continue
            seen_states.add(state_tuple)

            triggered = execute_Tr(state_tuple[0], state_tuple[1], state_tuple[2])
            reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)

            samples_with_scores.append((state_tuple, reward, sim, triggered))

        samples_with_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_scores[:num_samples]


def load_path_data(file_path):
    """加载screening生成的Path样本"""
    path_data = []

    if not os.path.exists(file_path):
        print(f"文件不存在: {file_path}")
        return path_data

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
            for line in lines[2:]:
                parts = line.strip().split("\t")
                if parts:
                    values = parts[0].split()
                    if len(values) >= 3:
                        state = (int(values[0]), int(values[1]), int(values[2]))
                        path_data.append(state)
    except Exception as e:
        print(f"读取文件出错 {file_path}: {e}")

    return path_data


# === DQN网络 ===
class DQN(nn.Module):
    """Q网络"""

    def __init__(self, state_dim, action_dim, hidden_dims=[128, 64]):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, hidden_dims[0])
        self.fc2 = nn.Linear(hidden_dims[0], hidden_dims[1])
        self.fc3 = nn.Linear(hidden_dims[1], action_dim)
        self.dropout = nn.Dropout(0.1)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = self.dropout(x)
        x = torch.relu(self.fc2(x))
        return self.fc3(x)


# === DQN Agent with PER ===
class DQNAgentWithPER:
    """DQN智能体(带优先级经验回放)"""

    def __init__(self, state_dim, action_dim, replay_buffer,
                 gamma=0.99, epsilon=1.0, epsilon_decay=0.995,
                 epsilon_min=0.1, learning_rate=0.001, alpha=0.6, beta=0.4):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer
        self.alpha = alpha
        self.beta = beta

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        """
        解码动作索引为具体的增量
        30个动作, 含义: 
        - weather: +/-1, 0(x2)
        - time_period: +/-1, 0(x2)
        - z: +/-12(20%), +/-6(10%), +/-3(5%), 0(x2)
        """
        delta_values_weather_time = [1, 0, 0, -1]  # 4个值
        delta_values_z = [12, 6, 3, 0, 0, -3, -6, -12]  # 8个值

        dim = action_idx // 10
        delta_idx = action_idx % 10

        if dim == 0:  # weather
            if delta_idx >= 4:
                delta_idx = 3
            return (delta_values_weather_time[delta_idx], 0, 0)
        elif dim == 1:  # time_period
            if delta_idx >= 4:
                delta_idx = 3
            return (0, delta_values_weather_time[delta_idx], 0)
        elif dim == 2:  # z
            if delta_idx >= 8:
                delta_idx = 7
            return (0, 0, delta_values_z[delta_idx])
        return (0, 0, 0)  # 默认返回

    def act(self, state_norm, legal_actions=None):
        """选择动作(epsilon-greedy)"""
        if legal_actions is None:
            legal_actions = list(range(self.action_dim))

        if not legal_actions:
            return None

        if random.random() < self.epsilon:
            return random.choice(legal_actions)

        state_tensor = torch.tensor(state_norm, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)[0]

        legal_q_values = q_values[legal_actions]
        best_legal_idx = torch.argmax(legal_q_values).item()
        return legal_actions[best_legal_idx]

    def get_legal_actions(self, state):
        """获取合法动作列表(边界检查)"""
        legal_actions = []

        for action_idx in range(self.action_dim):
            dw, dt, dz = self.decode_action(action_idx)

            next_weather = state[0] + dw
            next_time = state[1] + dt
            next_z = state[2] + dz

            if (MIN_X <= next_weather <= MAX_X and
                    MIN_Y <= next_time <= MAX_Y and
                    MIN_Z <= next_z <= MAX_Z):
                legal_actions.append(action_idx)

        return legal_actions

    def store_transition(self, state, action, reward, next_state, done):
        """存储转换(自动归一化)"""
        # 归一化
        state_norm = normalize_state(state)
        next_state_norm = normalize_state(next_state)

        state_tensor = torch.tensor(state_norm, dtype=torch.float32).unsqueeze(0).to(device)
        next_state_tensor = torch.tensor(next_state_norm, dtype=torch.float32).unsqueeze(0).to(device)

        with torch.no_grad():
            q_values = self.model(state_tensor)
            next_q_values = self.target_model(next_state_tensor)
            max_next_q = next_q_values.max(1)[0]
            target_q = reward + (self.gamma * max_next_q * (1 - done))
            td_error = abs(q_values[0][action].item() - target_q.item())

        self.replay_buffer.append((state_tensor, action, reward, next_state_tensor, done, td_error))
        return td_error

    def train(self, batch_size=32):
        """训练一步"""
        if len(self.replay_buffer) < batch_size:
            return 0.0

        batch, batch_indices, probabilities = self.replay_buffer.sample(batch_size, alpha=self.alpha)

        if not batch:
            return 0.0

        states, actions, rewards, next_states, dones, _ = zip(*batch)

        weights = (len(self.replay_buffer) * probabilities) ** (-self.beta)
        weights = weights / weights.max()
        weights = torch.tensor(weights, dtype=torch.float32).to(device)

        states = torch.cat(states).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.cat(next_states).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q = self.target_model(next_states).max(1)[0].detach()
        target_q = rewards + (self.gamma * next_max_q * (1 - dones))

        td_errors = current_q - target_q
        weighted_loss = (td_errors.pow(2) * weights).mean()

        self.optimizer.zero_grad()
        weighted_loss.backward()
        torch.nn.utils.clip_grad_norm_(self.model.parameters(), 1.0)
        self.optimizer.step()

        new_priorities = abs(td_errors.detach().cpu().numpy())
        self.replay_buffer.update_priorities(batch_indices, new_priorities)

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

        return weighted_loss.item()

    def update_target_model(self):
        """更新目标网络"""
        self.target_model.load_state_dict(self.model.state_dict())


# === 独立路径训练 ===
def generate_and_train_for_individual_paths(path_documents, repeats=5, batch_size=32, run_id=1):
    """独立训练每条路径(共享经验池)"""
    state_dim = 3
    action_dim = 30

    shared_replay_buffer = SharedExperienceReplay(capacity=20000)
    agent = DQNAgentWithPER(state_dim, action_dim, shared_replay_buffer)

    total_cumulative_reward = 0
    path_rewards = {}

    print(f"\n=== 第 {run_id}/20 轮训练开始(共享经验池) ===")
    start_time = time.time()

    SAMPLES_PER_BATCH = 50
    NUM_BATCHES = 4
    STEPS_PER_SAMPLE = 3

    for path_idx in range(len(targetPaths)):
        path_id = path_idx + 1
        print(f"\n{'=' * 60}")
        print(f"路径 {path_id}/{NUM_PATHS}")
        print(f"{'=' * 60}")

        file_path = os.path.join(path_documents, f"path{path_id}_individual.txt")
        if not os.path.exists(file_path):
            print(f"  跳过: 样本文件不存在 {file_path}")
            continue

        path_data = load_path_data(file_path)
        if not path_data:
            print(f"  跳过: 路径 {path_id} 无有效样本")
            continue

        target_path = targetPaths[path_idx]

        if path_idx not in path_rewards:
            path_rewards[path_idx] = 0

        for repeat_idx in range(repeats):
            print(f"\n  轮次 {repeat_idx + 1}/{repeats} 开始")

            for batch_idx in range(NUM_BATCHES):
                batch_start = batch_idx * SAMPLES_PER_BATCH
                batch_end = min(batch_start + SAMPLES_PER_BATCH, len(path_data))

                print(f"    批次 {batch_idx + 1}/{NUM_BATCHES} (样本 {batch_start}-{batch_end})")

                for sample_idx in range(batch_start, batch_end):
                    state = path_data[sample_idx]
                    prev_state = None
                    prev_triggered = None

                    for step in range(STEPS_PER_SAMPLE):
                        legal_actions = agent.get_legal_actions(state)

                        if not legal_actions:
                            break

                        # 归一化状态用于网络输入
                        state_norm = normalize_state(state)
                        action = agent.act(state_norm, legal_actions)
                        if action is None:
                            break

                        dw, dt, dz = agent.decode_action(action)

                        next_state = (
                            int(np.clip(state[0] + dw, MIN_X, MAX_X)),
                            int(np.clip(state[1] + dt, MIN_Y, MAX_Y)),
                            int(np.clip(state[2] + dz, MIN_Z, MAX_Z))
                        )

                        triggered = execute_Tr(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered,
                                                prev_triggered, prev_state)
                        done = (step == STEPS_PER_SAMPLE - 1)

                        # 存储转换
                        agent.store_transition(state, action, reward, next_state, done)

                        prev_state = state
                        prev_triggered = triggered
                        state = next_state

                        total_cumulative_reward += reward
                        path_rewards[path_idx] += reward

                if len(agent.replay_buffer) >= batch_size:
                    loss = agent.train(batch_size)
                    print(f"      批次 {batch_idx + 1} 完成, 损失: {loss:.4f}")

                if (batch_idx + 1) % 2 == 0:
                    agent.update_target_model()
                    print(f"      目标网络已更新(批次 {batch_idx + 1})")

        print(f"\n路径 {path_id} 完成, 累计奖励: {path_rewards[path_idx]:.2f}")
        print(f"共享缓冲区大小: {len(shared_replay_buffer)}")

    training_time = time.time() - start_time
    print(f"\n=== 第 {run_id}/20 轮完成, 用时: {training_time:.2f} 秒 ===")

    return agent, shared_replay_buffer, total_cumulative_reward, path_rewards, training_time


# === Excel导出 ===
def create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir):
    """创建综合Excel报告"""
    os.makedirs(output_dir, exist_ok=True)

    similar_group_paths = [idx + 1 for idx in similar_group]
    isolated_group_paths = [idx + 1 for idx in isolated_group]

    wb = Workbook()

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    header_color = "4472C4"
    similar_group_color = "E2EFDA"
    isolated_group_color = "FCE4D6"
    stats_color = "FFF2CC"

    # === Sheet 1: 路径相似度统计 ===
    ws_paths = wb.active
    ws_paths.title = "路径相似度统计"

    path_headers = ['路径ID', '分组类型'] + [f'第{i}轮' for i in range(1, 21)] + \
                   ['平均相似度', '最高相似度', '最低相似度', '标准差']

    for col, header in enumerate(path_headers, 1):
        cell = ws_paths.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_paths.row_dimensions[1].height = 30

    for path_id in range(1, NUM_PATHS + 1):
        row = path_id + 1

        if path_id in similar_group_paths:
            group_type = "高相关路径组"
            row_color = similar_group_color
        elif path_id in isolated_group_paths:
            group_type = "低相关路径组"
            row_color = isolated_group_color
        else:
            group_type = "未分组"
            row_color = "FFFFFF"

        cell = ws_paths.cell(row=row, column=1, value=f"路径 {path_id}")
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        cell.border = thin_border

        cell = ws_paths.cell(row=row, column=2, value=group_type)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        cell.border = thin_border

        path_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            sim = run_data['path_similarities'].get(path_id, {}).get('avg_similarity', 0.0)
            path_similarities.append(sim)

            cell = ws_paths.cell(row=row, column=3 + run_idx, value=round(sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        stats_values = [
            round(np.mean(path_similarities), 4),
            round(np.max(path_similarities), 4),
            round(np.min(path_similarities), 4),
            round(np.std(path_similarities), 4)
        ]

        for i, value in enumerate(stats_values):
            cell = ws_paths.cell(row=row, column=23 + i, value=value)
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
            cell.font = Font(bold=True, size=10)
            cell.border = thin_border

    ws_paths.column_dimensions['A'].width = 13
    ws_paths.column_dimensions['B'].width = 16
    for col in range(3, 23):
        ws_paths.column_dimensions[get_column_letter(col)].width = 10
    for col in range(23, 27):
        ws_paths.column_dimensions[get_column_letter(col)].width = 13

    # === Sheet 2: 分组统计 ===
    ws_groups = wb.create_sheet("分组统计")

    group_headers = ['组名', '包含路径'] + [f'第{i}轮' for i in range(1, 21)] + \
                    ['平均相似度', '标准差']

    for col, header in enumerate(group_headers, 1):
        cell = ws_groups.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_groups.row_dimensions[1].height = 30

    row = 2

    # 高相关路径组
    cell = ws_groups.cell(row=row, column=1, value="高相关路径组")
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    cell.border = thin_border

    cell = ws_groups.cell(row=row, column=2, value=','.join(map(str, similar_group_paths)))
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    cell.border = thin_border

    group_similarities = []
    for run_idx, run_data in enumerate(all_runs_data):
        group_sim = np.mean([
            run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
            for p in similar_group_paths
        ])
        group_similarities.append(group_sim)

        cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(group_sim, 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    cell = ws_groups.cell(row=row, column=23, value=round(np.mean(group_similarities), 4))
    cell.number_format = '0.0000'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border

    cell = ws_groups.cell(row=row, column=24, value=round(np.std(group_similarities), 4))
    cell.number_format = '0.0000'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border

    row += 1

    # 低相关路径组
    if isolated_group_paths:
        cell = ws_groups.cell(row=row, column=1, value="低相关路径组")
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        cell.border = thin_border

        cell = ws_groups.cell(row=row, column=2, value=','.join(map(str, isolated_group_paths)))
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        cell.border = thin_border

        isolated_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            iso_sim = np.mean([
                run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
                for p in isolated_group_paths
            ])
            isolated_similarities.append(iso_sim)

            cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(iso_sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        cell = ws_groups.cell(row=row, column=23, value=round(np.mean(isolated_similarities), 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border

        cell = ws_groups.cell(row=row, column=24, value=round(np.std(isolated_similarities), 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border

    ws_groups.column_dimensions['A'].width = 16
    ws_groups.column_dimensions['B'].width = 22
    for col in range(3, 23):
        ws_groups.column_dimensions[get_column_letter(col)].width = 10
    ws_groups.column_dimensions[get_column_letter(23)].width = 14
    ws_groups.column_dimensions[get_column_letter(24)].width = 12

    # === Sheet 3: 详细样本数据 ===
    ws_samples = wb.create_sheet("详细样本数据")

    sample_headers = ['轮次', '路径ID', '样本ID', 'Weather', 'TimePeriod', 'Z', '相似度', '触发规则集']

    for col, header in enumerate(sample_headers, 1):
        cell = ws_samples.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_samples.row_dimensions[1].height = 30

    sample_row = 2
    for run_idx, run_data in enumerate(all_runs_data, 1):
        for path_id in range(1, NUM_PATHS + 1):
            samples = run_data['path_samples'].get(path_id, [])

            if path_id in similar_group_paths:
                path_color = similar_group_color
            elif path_id in isolated_group_paths:
                path_color = isolated_group_color
            else:
                path_color = "FFFFFF"

            for sample_idx, (state_tuple, reward, sim, triggered) in enumerate(samples, 1):
                weather, time_period, z = state_tuple
                triggered_str = ','.join(map(str, sorted(triggered)))

                cell = ws_samples.cell(row=sample_row, column=1, value=f"第{run_idx}轮")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=path_color, end_color=path_color, fill_type="solid")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=2, value=f"路径{path_id}")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=path_color, end_color=path_color, fill_type="solid")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=3, value=sample_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                for col_offset, value in enumerate([weather, time_period, z]):
                    cell = ws_samples.cell(row=sample_row, column=4 + col_offset, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=7, value=round(sim, 4))
                cell.number_format = '0.0000'
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=8, value=f"{{{triggered_str}}}")
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border

                sample_row += 1

    sample_widths = [13, 13, 11, 10, 12, 8, 12, 45]
    for i, width in enumerate(sample_widths, 1):
        ws_samples.column_dimensions[get_column_letter(i)].width = width

    # === Sheet 4: 运行统计摘要 ===
    ws_summary = wb.create_sheet("运行统计摘要")

    summary_headers = ['轮次', '训练时间(秒)', '整体平均相似度', '最高相似度',
                       '最低相似度', '高相关组平均', '低相关组平均', '共享缓冲区大小']

    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_summary.row_dimensions[1].height = 30

    for run_idx, run_data in enumerate(all_runs_data, 1):
        row = run_idx + 1

        # 计算分组平均相似度
        high_group_avg = np.mean([
            run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
            for p in similar_group_paths
        ])

        low_group_avg = 0.0
        if isolated_group_paths:
            low_group_avg = np.mean([
                run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
                for p in isolated_group_paths
            ])

        values = [
            f"第{run_idx}轮",
            round(run_data['training_time'], 2),
            round(run_data['overall_avg_similarity'], 4),
            round(run_data['max_similarity'], 4),
            round(run_data['min_similarity'], 4),
            round(high_group_avg, 4),
            round(low_group_avg, 4),
            20000  # 缓冲区容量
        ]

        for col, value in enumerate(values, 1):
            cell = ws_summary.cell(row=row, column=col, value=value)
            if col == 1:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 2:
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif 3 <= col <= 7:
                cell.number_format = '0.0000'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # 统计行
    stat_row = len(all_runs_data) + 2
    stat_labels = ['总计/平均', '', '', '', '', '', '', '']

    for col, label in enumerate(stat_labels, 1):
        cell = ws_summary.cell(row=stat_row, column=col, value=label)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    stat_row += 1

    # 计算统计数据
    training_times = [r['training_time'] for r in all_runs_data]
    overall_avgs = [r['overall_avg_similarity'] for r in all_runs_data]
    max_sims = [r['max_similarity'] for r in all_runs_data]
    min_sims = [r['min_similarity'] for r in all_runs_data]

    high_group_avgs = []
    low_group_avgs = []
    for run_data in all_runs_data:
        high_avg = np.mean([
            run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
            for p in similar_group_paths
        ])
        high_group_avgs.append(high_avg)

        if isolated_group_paths:
            low_avg = np.mean([
                run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
                for p in isolated_group_paths
            ])
            low_group_avgs.append(low_avg)

    stat_values = [
        '统计值',
        round(np.sum(training_times), 2),
        round(np.mean(overall_avgs), 4),
        round(np.max(max_sims), 4),
        round(np.min(min_sims), 4),
        round(np.mean(high_group_avgs), 4),
        round(np.mean(low_group_avgs), 4) if low_group_avgs else 0.0,
        20000
    ]

    for col, value in enumerate(stat_values, 1):
        cell = ws_summary.cell(row=stat_row, column=col, value=value)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        if col == 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col == 2:
            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif 3 <= col <= 7:
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    summary_widths = [13, 16, 18, 14, 14, 16, 16, 14]
    for i, width in enumerate(summary_widths, 1):
        ws_summary.column_dimensions[get_column_letter(i)].width = width

    output_path = os.path.join(output_dir, "20轮训练结果.xlsx")
    wb.save(output_path)
    print(f"\n综合Excel报告已生成: {output_path}")
    print(f"  包含4个工作表: 路径相似度统计, 分组统计, 详细样本数据, 运行统计摘要")


def run_20_times_training():
    """运行20轮训练"""
    model_path_base = r"D:\Experiment\CNN\DQNNEW\saved_models_new_vars"
    path_documents = r"D:\Experiment\CNN\DQNNEW\path_samples_individual"
    output_dir = r"D:\Experiment\CNN\ComparisonExperiment2\excel_reports_new_vars"

    os.makedirs(model_path_base, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    similar_group, isolated_group = group_paths_by_similarity(targetPaths)

    similar_group_display = [idx + 1 for idx in similar_group]
    isolated_group_display = [idx + 1 for idx in isolated_group]

    print("=" * 60)
    print("20轮训练 - 共享经验池版本")
    print("=" * 60)
    print(f"\n路径总数: {NUM_PATHS} 条")
    print(f"\n自动分组结果:")
    print(f"  高相关路径组: {similar_group_display}")
    print(f"  低相关路径组: {isolated_group_display}")
    print(f"\n取值范围:")
    print(f"  X (weather): {MIN_X}-{MAX_X}")
    print(f"  Y (time_period): {MIN_Y}-{MAX_Y}")
    print(f"  Z: {MIN_Z}-{MAX_Z}")
    print(f"\n动作空间 (30个):")
    print(f"  weather: +/-1, 0(x2)")
    print(f"  time_period: +/-1, 0(x2)")
    print(f"  z: +/-12, +/-6, +/-3, 0(x2)")
    print("\n" + "=" * 60)

    all_runs_data = []
    total_start_time = time.time()

    for run_id in range(1, 21):
        print(f"\n{'=' * 60}")
        print(f"开始第 {run_id}/20 轮训练")
        print(f"{'=' * 60}")

        print(f"[第{run_id}轮] 生成样本...")
        generate_samples_for_all_paths(num_candidates=2000, top_k=200, run_id=run_id)

        print(f"[第{run_id}轮] 开始训练...")
        agent, shared_buffer, total_cumulative_reward, path_rewards, training_time = \
            generate_and_train_for_individual_paths(path_documents, repeats=5,
                                                    batch_size=32, run_id=run_id)

        model_path = os.path.join(model_path_base, f"trained_model_run_{run_id}.pth")
        torch.save({
            'model_state_dict': agent.model.state_dict(),
            'optimizer_state_dict': agent.optimizer.state_dict(),
            'epsilon': agent.epsilon,
            'run_id': run_id,
            'normalized': True,
            'value_ranges': {
                'weather': [MIN_X, MAX_X],
                'time_period': [MIN_Y, MAX_Y],
                'z': [MIN_Z, MAX_Z]
            }
        }, model_path)
        print(f"[第{run_id}轮] 模型已保存: {model_path}")

        run_data = {
            'run_id': run_id,
            'training_time': training_time,
            'total_reward': total_cumulative_reward,
            'path_rewards': path_rewards,
            'path_similarities': {},
            'path_samples': {}
        }

        all_similarities = []
        for path_idx in range(len(targetPaths)):
            target_path = targetPaths[path_idx]
            high_reward_samples = shared_buffer.get_high_reward_samples(target_path, num_samples=20)

            if high_reward_samples:
                similarities = [sim for _, _, sim, _ in high_reward_samples]
                run_data['path_similarities'][path_idx + 1] = {
                    'avg_similarity': np.mean(similarities),
                    'max_similarity': np.max(similarities),
                    'min_similarity': np.min(similarities),
                    'sample_count': len(similarities)
                }
                run_data['path_samples'][path_idx + 1] = high_reward_samples
                all_similarities.extend(similarities)
            else:
                run_data['path_similarities'][path_idx + 1] = {
                    'avg_similarity': 0.0,
                    'max_similarity': 0.0,
                    'min_similarity': 0.0,
                    'sample_count': 0
                }
                run_data['path_samples'][path_idx + 1] = []

        if all_similarities:
            run_data['overall_avg_similarity'] = np.mean(all_similarities)
            run_data['max_similarity'] = np.max(all_similarities)
            run_data['min_similarity'] = np.min(all_similarities)
        else:
            run_data['overall_avg_similarity'] = 0.0
            run_data['max_similarity'] = 0.0
            run_data['min_similarity'] = 0.0

        all_runs_data.append(run_data)

        print(f"[第{run_id}轮] 完成! 整体平均相似度: {run_data['overall_avg_similarity']:.4f}")
        print(f"{'=' * 60}\n")

    total_time = time.time() - total_start_time

    print("\n正在生成综合Excel报告...")
    create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir)

    print("\n" + "=" * 60)
    print("20轮训练全部完成!")
    print("=" * 60)
    print(f"总耗时: {total_time:.2f} 秒 ({total_time / 60:.2f} 分钟)")
    print(f"平均每轮耗时: {total_time / 20:.2f} 秒")
    print(f"\n所有结果已保存到: {output_dir}")
    print("=" * 60)


if __name__ == "__main__":
    run_20_times_training()