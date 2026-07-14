import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# X, Y, Z 范围 - 使用提供的范围
X_MIN = 2
X_MAX = 100
Y_MIN = 1
Y_MAX = 150
Z_MIN = 1
Z_MAX = 200


# === 归一化/反归一化 ===
def normalize_state(state):
    """将状态归一化到(0-1)范围"""
    x, y, z = state
    norm_x = (x - X_MIN) / (X_MAX - X_MIN)
    norm_y = (y - Y_MIN) / (Y_MAX - Y_MIN)
    norm_z = (z - Z_MIN) / (Z_MAX - Z_MIN)
    return (norm_x, norm_y, norm_z)


def denormalize_state(norm_state):
    """将归一化状态还原到原始范围"""
    norm_x, norm_y, norm_z = norm_state
    x = int(round(norm_x * (X_MAX - X_MIN) + X_MIN))
    y = int(round(norm_y * (Y_MAX - Y_MIN) + Y_MIN))
    z = int(round(norm_z * (Z_MAX - Z_MIN) + Z_MIN))
    # 边界保护
    x = max(X_MIN, min(X_MAX, x))
    y = max(Y_MIN, min(Y_MAX, y))
    z = max(Z_MIN, min(Z_MAX, z))
    return (x, y, z)


def is_valid_state(state):
    """检查状态是否有效"""
    x, y, z = state
    return (X_MIN <= x <= X_MAX and
            Y_MIN <= y <= Y_MAX and
            Z_MIN <= z <= Z_MAX)


def clip_state(state):
    """裁剪状态到有效范围"""
    x, y, z = state
    return (
        max(X_MIN, min(X_MAX, x)),
        max(Y_MIN, min(Y_MAX, y)),
        max(Z_MIN, min(Z_MAX, z))
    )


# === 增强标准指标收集器 ===
class EnhancedStandardMetricsCollector:
    def __init__(self, experiment_name="Enhanced_Standard_DQN"):
        self.experiment_name = experiment_name
        self.start_time = None
        self.end_time = None

        # 基础指标
        self.total_reward = 0
        self.td_errors = []
        self.final_output_similarities = []
        self.action_improvements = []
        self.total_memory_usage = 0
        self.memory_check_count = 0
        self.step_count = 0

        # Episode级指标
        self.episode_rewards = []
        self.episode_similarities = []
        self.episode_td_errors = []
        self.episode_epsilon_values = []
        self.episode_memory_usage = []

        # 路径类型指标
        self.similar_paths_performance = []
        self.isolated_paths_performance = []

        # 里程碑数据
        self.milestone_data = {}

        # 收敛检测
        self.convergence_window = 20
        self.convergence_threshold = 0.02
        self.convergence_detected_episode = None

        # 样本效率指标
        self.sample_efficiency_data = []
        self.performance_milestones = [0.6, 0.7, 0.75, 0.8]

        # 学习曲线特征
        self.learning_curve_characteristics = {}
        self.early_vs_late_performance = {}

        # X, Y, Z坐标统计
        self.xyz_coordinate_stats = {
            'x_values': [],
            'y_values': [],
            'z_values': [],
            'x_distribution': {},
            'y_distribution': {},
            'z_distribution': {}
        }

    def reset(self):
        """重置所有指标"""
        self.start_time = None
        self.end_time = None
        self.total_reward = 0
        self.td_errors = []
        self.final_output_similarities = []
        self.action_improvements = []
        self.total_memory_usage = 0
        self.memory_check_count = 0
        self.step_count = 0
        self.episode_rewards = []
        self.episode_similarities = []
        self.episode_td_errors = []
        self.episode_epsilon_values = []
        self.episode_memory_usage = []
        self.similar_paths_performance = []
        self.isolated_paths_performance = []
        self.milestone_data = {}
        self.convergence_detected_episode = None
        self.sample_efficiency_data = []
        self.learning_curve_characteristics = {}
        self.early_vs_late_performance = {}
        self.xyz_coordinate_stats = {
            'x_values': [],
            'y_values': [],
            'z_values': [],
            'x_distribution': {},
            'y_distribution': {},
            'z_distribution': {}
        }

    def start_training(self):
        self.start_time = time.time()

    def end_training(self):
        self.end_time = time.time()

    def record_xyz_coordinates(self, x_val, y_val, z_val):
        """记录X, Y, Z坐标值"""
        self.xyz_coordinate_stats['x_values'].append(x_val)
        self.xyz_coordinate_stats['y_values'].append(y_val)
        self.xyz_coordinate_stats['z_values'].append(z_val)

        self.xyz_coordinate_stats['x_distribution'][x_val] = self.xyz_coordinate_stats['x_distribution'].get(x_val, 0) + 1
        self.xyz_coordinate_stats['y_distribution'][y_val] = self.xyz_coordinate_stats['y_distribution'].get(y_val, 0) + 1
        self.xyz_coordinate_stats['z_distribution'][z_val] = self.xyz_coordinate_stats['z_distribution'].get(z_val, 0) + 1

    def record_step_metrics(self, reward, td_error, triggered, target_path, x_coord=None, y_coord=None, z_coord=None):
        """记录步骤级指标"""
        self.step_count += 1
        self.total_reward += reward
        self.td_errors.append(td_error)

        if x_coord is not None and y_coord is not None and z_coord is not None:
            self.record_xyz_coordinates(x_coord, y_coord, z_coord)

        try:
            process = psutil.Process(os.getpid())
            current_memory = process.memory_info().rss / 1024 / 1024
            self.total_memory_usage += current_memory
            self.memory_check_count += 1
        except:
            pass

    def record_episode_metrics(self, episode, episode_reward, avg_similarity, avg_td_error, epsilon, path_group="similar"):
        """记录episode级指标"""
        self.episode_rewards.append(episode_reward)
        self.episode_similarities.append(avg_similarity)
        self.episode_td_errors.append(avg_td_error)
        self.episode_epsilon_values.append(epsilon)

        try:
            process = psutil.Process(os.getpid())
            current_memory = process.memory_info().rss / 1024 / 1024
            self.episode_memory_usage.append(current_memory)
        except:
            self.episode_memory_usage.append(0)

        if path_group == "similar":
            self.similar_paths_performance.append({
                'episode': episode,
                'reward': episode_reward,
                'similarity': avg_similarity,
                'td_error': avg_td_error
            })
        else:
            self.isolated_paths_performance.append({
                'episode': episode,
                'reward': episode_reward,
                'similarity': avg_similarity,
                'td_error': avg_td_error
            })

        if episode in [50, 100, 150, 200, 250, 300, 400, 450, 500]:
            self.milestone_data[episode] = {
                'avg_reward': np.mean(self.episode_rewards[-10:]) if len(self.episode_rewards) >= 10 else episode_reward,
                'avg_similarity': avg_similarity,
                'avg_td_error': avg_td_error,
                'epsilon': epsilon,
                'total_steps': self.step_count
            }

        self._check_convergence(episode)
        self._check_performance_milestones(episode, avg_similarity)

    def _check_convergence(self, episode):
        """检查是否收敛"""
        if len(self.episode_similarities) >= self.convergence_window and self.convergence_detected_episode is None:
            recent_similarities = self.episode_similarities[-self.convergence_window:]
            if np.std(recent_similarities) < self.convergence_threshold:
                self.convergence_detected_episode = episode

    def _check_performance_milestones(self, episode, similarity):
        """检查性能里程碑"""
        for milestone in self.performance_milestones:
            if similarity >= milestone and not any(data[1] == milestone for data in self.sample_efficiency_data):
                self.sample_efficiency_data.append((episode, milestone, self.step_count))

    def record_final_output_sample(self, triggered, target_path):
        """记录最终输出样本的相似度"""
        if len(triggered | target_path) > 0:
            similarity = len(triggered & target_path) / len(triggered | target_path)
        else:
            similarity = 0.0
        self.final_output_similarities.append(similarity)

    def record_action_improvement(self, current_reward, prev_reward):
        """记录动作改进"""
        if prev_reward is not None:
            improvement = current_reward - prev_reward
            self.action_improvements.append(1 if improvement > 0 else 0)


# 全局指标收集器
enhanced_standard_metrics = EnhancedStandardMetricsCollector("Enhanced_Standard_DQN_No_Priority")


# === reward function ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward


def execute_Tr(x, y, z):
    """执行变异分支，返回触发的规则集合 - 85个分支版本"""
    triggered = set()

    # --- 分支 1-11 (原 energy_y * energy_z / (energy_x + 1) > 140 的变异) ---
    if ((y * z) / (x + 1) > 140) != ((y * y) / (x + 1) > 140): triggered.add(1)
    if ((y * z) / (x + 1) > 140) != ((z * z) / (x + 1) > 140): triggered.add(2)
    if ((y * z) / (x + 1) > 140) != ((y * x) / (x + 1) > 140): triggered.add(3)
    if ((y * z) / (x + 1) > 140) != ((y * z) / (x + 3) > 140): triggered.add(4)
    if ((y * z) / (x + 1) > 140) != ((y * z) / (x - 1) > 140): triggered.add(5)
    if ((y * z) / (x + 1) > 140) != ((y * z * 2) / (y + 1) > 140): triggered.add(6)
    if ((y * z) / (x + 1) > 140) != ((y * z) / (x + 1) > 100): triggered.add(7)
    if ((y * z) / (x + 1) > 140) != ((y * z) / (x + 1) > 180): triggered.add(8)
    if ((y * z) / (x + 1) > 140) != ((y * z) / (x + 10) > 140): triggered.add(9)
    if ((y * z) / (x + 1) > 140) != ((y * z) / (x * 1) > 140): triggered.add(10)
    if ((y * z) / (x + 1) > 140) != ((y * 30) / (x + 1) > 140): triggered.add(11)

    # --- 分支 12-21 (原 (energy_z - energy_x) < 0.22 * energy_y 的变异) ---
    if ((z - x) < 0.22 * y) != ((z - x) < 0.22 * x): triggered.add(12)
    if ((z - x) < 0.22 * y) != ((z - x) < 0.22 * z): triggered.add(13)
    if ((z - x) < 0.22 * y) != ((z - x) < 0.32 * y): triggered.add(14)
    if ((z - x) < 0.22 * y) != ((z - x) < 0.12 * y): triggered.add(15)
    if ((z - x) < 0.22 * y) != ((z * 2 - x) < 0.22 * y): triggered.add(16)
    if ((z - x) < 0.22 * y) != ((z - x * 1.2) < 0.22 * y): triggered.add(17)
    if ((z - x) < 0.22 * y) != ((z + x) < 0.22 * y): triggered.add(18)
    if ((z - x) < 0.22 * y) != ((z - 20) < 0.22 * y): triggered.add(19)
    if ((z - x) < 0.22 * y) != ((90 - x) < 0.22 * y): triggered.add(20)
    if ((z - x) < 0.22 * y) != ((z - x) < 0.4 * y): triggered.add(21)

    # --- 分支 22-32 (原 (energy_x^3 + energy_y^3) < energy_z^2 的变异) ---
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2.7 + y ** 3) < z ** 2): triggered.add(22)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 2.6) < z ** 2): triggered.add(23)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 1.8): triggered.add(24)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 - y ** 3) < z ** 2): triggered.add(25)
    if ((x ** 3 + y ** 3) < z ** 2) != ((y ** 3 + y ** 3) < z ** 2): triggered.add(26)
    if ((x ** 3 + y ** 3) < z ** 2) != ((z ** 3 + y ** 3) < z ** 2): triggered.add(27)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + x ** 3) < z ** 2): triggered.add(28)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + z ** 3) < z ** 2): triggered.add(29)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < x ** 2): triggered.add(30)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < y ** 2): triggered.add(31)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 2.5): triggered.add(32)

    # --- 分支 33-42 (原 x/(y+0.01)>5 and y/(z+0.01)<0.2 的变异) ---
    if ((x / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.2) != ((x / (z + 0.01)) > 5 and (y / (z + 0.01)) < 0.2): triggered.add(33)
    if ((x / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.2) != ((x / (x + 0.01)) > 5 and (y / (z + 0.01)) < 0.2): triggered.add(34)
    if ((x / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.2) != ((z / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.2): triggered.add(35)
    if ((x / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.2) != ((y / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.2): triggered.add(36)
    if ((x / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.2) != ((x / (y + 0.01)) > 5 and (z / (z + 0.01)) < 0.2): triggered.add(37)
    if ((x / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.2) != ((x / (y + 0.01)) > 5 and (x / (z + 0.01)) < 0.2): triggered.add(38)
    if ((x / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.2) != ((x / (y + 0.01)) > 5 and (y / (y + 0.01)) < 0.2): triggered.add(39)
    if ((x / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.2) != ((x / (y + 0.01)) > 5 and (y / (x + 0.01)) < 0.2): triggered.add(40)
    if ((x / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.2) != ((x / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.15): triggered.add(41)
    if ((x / (y + 0.01)) > 5 and (y / (z + 0.01)) < 0.2) != ((x / (y + 0.01)) > 7 and (y / (z + 0.01)) < 0.2): triggered.add(42)

    # --- 分支 43-52 (原 abs(x-y)>16 and abs(y-z)>18 and abs(x-z)<9 的变异) ---
    if (abs(x - y) > 16 and abs(y - z) > 18 and abs(x - z) < 9) != (abs(x * 1.2 - y) > 16 and abs(y - z) > 18 and abs(x - z) < 9): triggered.add(43)
    if (abs(x - y) > 16 and abs(y - z) > 18 and abs(x - z) < 9) != (abs(x - y) > 16 and abs(y * 2 - z) > 18 and abs(x - z) < 9): triggered.add(44)
    if (abs(x - y) > 16 and abs(y - z) > 18 and abs(x - z) < 9) != (abs(x - y) > 19 and abs(y - z) > 18 and abs(y - z) < 9): triggered.add(45)
    if (abs(x - y) > 16 and abs(y - z) > 18 and abs(x - z) < 9) != (abs(x - y) > 16 and abs(x - z) > 18 and abs(x - z) < 9): triggered.add(46)
    if (abs(x - y) > 16 and abs(y - z) > 18 and abs(x - z) < 9) != (abs(x - y) > 16 and abs(y - z) > 40 and abs(x - z) < 9): triggered.add(47)
    if (abs(x - y) > 16 and abs(y - z) > 18 and abs(x - z) < 9) != (abs(x - y) > 16 and abs(y - z) > 18 and abs(x * 2 - z) < 9): triggered.add(48)
    if (abs(x - y) > 16 and abs(y - z) > 18 and abs(x - z) < 9) != (abs(x - y) > 16 and abs(y - z * 0.2) > 18 and abs(x - z) < 9): triggered.add(49)
    if (abs(x - y) > 16 and abs(y - z) > 18 and abs(x - z) < 9) != (abs(x - y) > 16 and abs(y - z) > 18 and abs(x * 1.5 - z) < 9): triggered.add(50)
    if (abs(x - y) > 16 and abs(y - z) > 18 and abs(x - z) < 9) != (abs(x - y) > 16 and abs(y - z) > 18 and abs(x - z * 0.87) < 9): triggered.add(51)
    if (abs(x - y) > 16 and abs(y - z) > 18 and abs(x - z) < 9) != (abs(x - y) > 16 and abs(y - z) > 18 and abs(x - z) < 7.8): triggered.add(52)

    # --- 分支 53-63 (原 (x>95 or x<5) and (y>90 or y<3) and (z>85 or z<2) 的变异) ---
    if ((x > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)) != ((x > 95 or x < 5) and (y * y > 90 or y < 3) and (z > 85 or z < 2)): triggered.add(53)
    if ((x > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)) != ((x > 95 or x < 5) and (y * x > 90 or y < 3) and (z > 85 or z < 2)): triggered.add(54)
    if ((x > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)) != ((x > 95 or x < 5) and (y * z > 90 or y < 3) and (z > 85 or z < 2)): triggered.add(55)
    if ((x > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)) != ((x > 95 or x < 5) and (y * 80 > 90 or y < 3) and (z > 85 or z < 2)): triggered.add(56)
    if ((x > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)) != ((x * y > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)): triggered.add(57)
    if ((x > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)) != ((x * x > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)): triggered.add(58)
    if ((x > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)) != ((x * z > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)): triggered.add(59)
    if ((x > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)) != ((x * 50 > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)): triggered.add(60)
    if ((x > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)) != ((x > 95 or x < 5) and (y * 40 > 90 or y < 3) and (z > 85 or z < 2)): triggered.add(61)
    if ((x > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)) != ((x > 95 or x < 5) and (y > 90 or y < 3) and (z * z > 85 or z < 2)): triggered.add(62)
    if ((x > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)) != ((x * 40 > 95 or x < 5) and (y > 90 or y < 3) and (z > 85 or z < 2)): triggered.add(63)

    # --- 分支 64-75 (原 x^0.7+y^0.7>z^0.9 and x+y+z<180 的变异) ---
    if (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180) != (x ** 0.6 + y ** 0.7 > z ** 0.9 and x + y + z < 180): triggered.add(64)
    if (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180) != (x ** 0.7 + y ** 0.7 > z ** 0.9 and z + y + z < 180): triggered.add(65)
    if (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180) != (x ** 0.7 + y ** 0.8 > z ** 0.9 and x + y + z < 180): triggered.add(66)
    if (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180) != (x ** 0.7 + y ** 0.7 > z ** 0.8 and x + y + z < 180): triggered.add(67)
    if (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180) != (x ** 0.7 + z ** 0.7 > z ** 0.9 and x + y + z < 180): triggered.add(68)
    if (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180) != (y ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180): triggered.add(69)
    if (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180) != (z ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180): triggered.add(70)
    if (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180) != (x ** 0.7 + x ** 0.7 > z ** 0.9 and x + y + z < 180): triggered.add(71)
    if (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180) != (x ** 0.7 + y ** 0.7 > x ** 0.9 and x + y + z < 180): triggered.add(72)
    if (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180) != (x ** 0.7 + y ** 0.7 > z ** 0.9 and y + y + z < 180): triggered.add(73)
    if (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180) != (x ** 0.7 + y ** 0.7 > z ** 0.9 and z + y + z < 180): triggered.add(74)
    if (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + y + z < 180) != (x ** 0.7 + y ** 0.7 > z ** 0.9 and x + x + z < 180): triggered.add(75)

    # --- 分支 76-85 (原 (x+y)^1.3<z^1.6 and x+y+z/3>35 的变异) ---
    if ((x + y) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35) != ((y + y) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35): triggered.add(76)
    if ((x + y) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35) != ((z + y) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35): triggered.add(77)
    if ((x + y) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35) != ((x + x) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35): triggered.add(78)
    if ((x + y) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35) != ((x + z) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35): triggered.add(79)
    if ((x + y) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35) != ((x + 20) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35): triggered.add(80)
    if ((x + y) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35) != ((x + y) ** 1 < z ** 1.6 and x + y + z / 3 > 35): triggered.add(81)
    if ((x + y) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35) != ((x + y) ** 1.3 < z ** 1.7 and x + y + z / 3 > 35): triggered.add(82)
    if ((x + y) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35) != ((x + y) ** 1.2 < z ** 1.6 and x + y + z / 3 > 35): triggered.add(83)
    if ((x + y) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35) != ((x + y) ** 1.3 < z ** 1.6 and y + y + z / 3 > 35): triggered.add(84)
    if ((x + y) ** 1.3 < z ** 1.6 and x + y + z / 3 > 35) != ((x + y) ** 1.3 < z ** 1.6 and x + y - z / 3 > 35): triggered.add(85)

    return triggered


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0


# === 目标路径定义 ===
targetPaths = [
    {2,3,6,8,9,11,12,13,15,16,18,19,20,25,48,50,65,68,71,73,74,78,79,80,81,82,83},  # A1
    {6,16,18,19,26,33,34,36,37,38,39,41,42,45,46,48,49,50,51,52,53,54,55,56,61},  # A2
    {2,3,4,6,8,9,11,12,13,15,16,18,20,25,48,50,68,71,73,78,79,80,81,82,83},  # A3
    {6,12,13,17,20,21,26,33,34,36,37,38,39,45,46,48,49,50,53,54,55,56,61},  # A4
    {1,16,18,19,20,25,45,46,47,48,50,64,65,68,71,73,74,78,79,80,81,82,83},  # A5
    {2,3,4,6,8,9,11,12,13,15,16,18,19,20,25,50,51,75,78,79,80,81,82,83},  # A6
    {12,13,17,20,26,33,34,36,37,38,39,45,46,48,49,50,51,57,58,59,60,63},  # A7
    {16,18,19,20,25,43,45,46,47,48,50,51,52,65,74,75,78,79,80,81,82,83},  # A8
    {16,18,19,20,25,43,45,46,47,48,50,64,68,71,73,77,78,79,80,81,82,83},  # A9
    {2,3,6,8,9,11,14,17,21,25,48,64,65,68,71,73,74,78,79,80,81,82,83},  # A10
    {1,7,12,13,16,18,19,20,25,50,51,65,68,71,73,74,78,79,80,81,82,83},  # A11
    {1,5,7,10,12,13,15,16,18,20,25,48,50,51,68,71,73,78,79,80,81},  # A12
    {12,13,17,20,26,33,34,36,37,38,39,45,46,48,50,58,59,60,63,84},  # A13
    {18,19,20,26,33,34,36,37,38,39,41,64,69,70,72,75,76,77,81,83},  # A14
    {16,18,19,20,26,33,34,35,36,37,38,39,41,67,68,71,78,79,80,84},  # A15
    {3,6,12,13,15,16,18,20,25,28,62,65,68,71,73,74,78,79,80,81},  # A16
    {16,18,19,20,26,33,34,36,37,38,39,67,68,71,78,79,80,84,85},  # A17
    {18,19,20,26,33,34,36,37,38,39,66,67,68,71,76,77,81,82,83},  # A18
    {2,6,24,27,28,29,30,31,33,34,36,37,39,58,59,60,63,85},  # A19
    {2,6,22,26,32,33,34,36,37,38,39,58,59,60,63,85},  # A20
    {12,13,14,17,21,26,44,45,46,47,48,49,50,84,85},  # A21
    {18,19,20,26,40,66,67,68,71,76,77,81,82,83},  # A22
    {3,23,25,28,32,35,53,54,55,56,61,85}  # A23
]


# === Path Similarity ===
def compute_path_similarity_matrix(paths):
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix


# === 路径分组 ===
def group_paths_by_similarity(paths):
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.mean(avg_sim_scores)

    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]
    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)

    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group


# === 标准经验回放(无优先级) ===
class StandardExperienceReplay:
    def __init__(self, capacity=10000):
        self.capacity = capacity
        self.buffer = deque(maxlen=self.capacity)

    def append(self, experience):
        """添加经验"""
        self.buffer.append(experience[:5])

    def sample(self, batch_size):
        """均匀采样"""
        if len(self.buffer) < batch_size:
            return [], [], None
        batch_indices = np.random.choice(len(self.buffer), batch_size, replace=False)
        batch = [self.buffer[idx] for idx in batch_indices]
        return batch, batch_indices, None

    def __len__(self):
        return len(self.buffer)

    def get_high_reward_samples(self, target_path, num_samples=20):
        """获取高奖励样本, 重新计算奖励"""
        if len(self.buffer) == 0:
            return []
        samples_with_recalculated_scores = []
        for experience in self.buffer:
            norm_state_tensor = experience[0]
            norm_state = tuple(norm_state_tensor.cpu().numpy().flatten())
            state_tuple = denormalize_state(norm_state)
            triggered = execute_Tr(state_tuple[0], state_tuple[1], state_tuple[2])
            new_reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_recalculated_scores.append((state_tuple, new_reward, sim, triggered))
        samples_with_recalculated_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_recalculated_scores[:num_samples]


def load_path_data(file_path):
    path_data = []
    if not os.path.exists(file_path):
        return path_data
    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines[2:]:
            parts = line.strip().split("\t")
            if parts:
                state = tuple(map(int, parts[0].split()))
                path_data.append(state)
    return path_data


# === DQN网络 ===
class DQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, 128)
        self.fc2 = nn.Linear(128, 64)
        self.fc3 = nn.Linear(64, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)


# === DQN Agent(标准版) ===
class StandardDQNAgent:
    def __init__(self, state_dim, action_dim, replay_buffer, gamma=0.99, epsilon=1.0,
                 epsilon_decay=0.995, epsilon_min=0.1, learning_rate=0.001):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        delta_values = [10, 5, 3, 2, 1, -1, -2, -3, -5, -10]
        dim = action_idx // 10
        delta_idx = action_idx % 10
        delta = delta_values[delta_idx]
        if dim == 0:
            return (delta, 0, 0)
        elif dim == 1:
            return (0, delta, 0)
        elif dim == 2:
            return (0, 0, delta)
        return (0, 0, 0)

    def act(self, state):
        """选择动作"""
        if random.random() < self.epsilon:
            return random.randrange(self.action_dim)
        norm_state = normalize_state(state)
        state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)
        return torch.argmax(q_values, dim=1).item()

    def store_transition(self, state, action, reward, next_state, done):
        """存储转换"""
        norm_state = normalize_state(state)
        norm_next_state = normalize_state(next_state)

        norm_state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
        norm_next_state_tensor = torch.tensor(norm_next_state, dtype=torch.float32).unsqueeze(0).to(device)

        with torch.no_grad():
            q_values = self.model(norm_state_tensor)
            next_q_values = self.target_model(norm_next_state_tensor)
            max_next_q_values = next_q_values.max(1)[0]
            target_q_values = reward + (self.gamma * max_next_q_values * (1 - done))
            td_error = torch.abs(q_values[0][action] - target_q_values).item()

        self.replay_buffer.append((norm_state_tensor, action, reward, norm_next_state_tensor, done))
        return td_error

    def train(self, batch_size=32):
        """训练一步"""
        if len(self.replay_buffer) < batch_size:
            return

        batch, batch_indices, _ = self.replay_buffer.sample(batch_size)
        states, actions, rewards, next_states, dones = zip(*batch)

        states = torch.tensor(np.array([s.cpu().numpy().flatten() for s in states]), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array([ns.cpu().numpy().flatten() for ns in next_states]),
                                   dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))

        loss = nn.MSELoss()(current_q_values, target_q_values)

        self.optimizer.zero_grad()
        loss.backward()
        self.optimizer.step()

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())


# === 样本生成 ===
def generate_samples_for_similar_paths(similar_group_indices, num_total=2000, top_k=200):
    def jaccard_similarity_local(a, b):
        if not a and not b: return 1.0
        return len(a & b) / len(a | b) if a | b else 0.0

    def compute_robustness(state, path):
        base = execute_Tr(state[0], state[1], state[2])
        if not base: return 0.0
        rob, neighbors = 0.0, 0
        for dx in [-1, 0, 1]:
            for dy in [-1, 0, 1]:
                for dz in [-1, 0, 1]:
                    if dx == dy == dz == 0: continue
                    neighbor_state = (state[0] + dx, state[1] + dy, state[2] + dz)
                    if not is_valid_state(neighbor_state): continue
                    neighbor = clip_state(neighbor_state)
                    n_trig = execute_Tr(neighbor[0], neighbor[1], neighbor[2])
                    if not n_trig: continue
                    rob += jaccard_similarity_local(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"enhanced_standard_path{path_id}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Enhanced Standard Path {path_id}\n")
            f.write("x y z\tScore\tSimilarity\tLengthDiff\tRobustness\n")
            for s in samples:
                x, y, z = s[0]
                f.write(f"{x} {y} {z}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_enhanced_standard"
    for path_idx in similar_group_indices:
        path = targetPaths[path_idx]
        samples = []
        attempts = 0
        while len(samples) < top_k and attempts < num_total * 5:
            attempts += 1
            state = (
                np.random.randint(X_MIN, X_MAX + 1),
                np.random.randint(Y_MIN, Y_MAX + 1),
                np.random.randint(Z_MIN, Z_MAX + 1)
            )
            triggered = execute_Tr(state[0], state[1], state[2])
            if not triggered: continue
            sim = jaccard_similarity_local(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            score = 0.55 * sim + 0.25 * len_diff + 0.2 * rob
            samples.append((state, score, sim, len_diff, rob))
        if samples:
            samples.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples[:top_k], base_dir=base_dir)


def generate_samples_for_isolated_paths_standard(agent_similar, isolated_group_indices, num_total=2000, top_k=200):
    def compute_q_value(state, agent):
        norm_state = normalize_state(state)
        state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = agent.model(state_tensor)
        return q_values.max().item()

    def compute_robustness(state, path):
        base = execute_Tr(state[0], state[1], state[2])
        if not base: return 0.0
        rob, neighbors = 0.0, 0
        for dx in [-1, 0, 1]:
            for dy in [-1, 0, 1]:
                for dz in [-1, 0, 1]:
                    if dx == dy == dz == 0: continue
                    neighbor_state = (state[0] + dx, state[1] + dy, state[2] + dz)
                    if not is_valid_state(neighbor_state): continue
                    neighbor = clip_state(neighbor_state)
                    n_trig = execute_Tr(neighbor[0], neighbor[1], neighbor[2])
                    if not n_trig: continue
                    rob += jaccard_similarity(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"enhanced_standard_path{path_id}_isolated.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Enhanced Standard Isolated Path {path_id}\n")
            f.write("x y z\tScore\tSimilarity\tLengthDiff\tRobustness\tQ_complement\n")
            for s in samples:
                x, y, z = s[0]
                f.write(f"{x} {y} {z}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\t{s[5]:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_enhanced_standard"
    for path_idx in isolated_group_indices:
        path = targetPaths[path_idx]
        samples_raw = []
        attempts = 0

        print(f"为路径 {path_idx + 1} 生成候选样本...")
        while len(samples_raw) < num_total and attempts < num_total * 5:
            attempts += 1
            state = (
                np.random.randint(X_MIN, X_MAX + 1),
                np.random.randint(Y_MIN, Y_MAX + 1),
                np.random.randint(Z_MIN, Z_MAX + 1)
            )
            triggered = execute_Tr(state[0], state[1], state[2])
            if not triggered: continue
            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            q_value = compute_q_value(state, agent_similar)
            samples_raw.append((state, sim, len_diff, rob, q_value))

        if not samples_raw:
            print(f"路径 {path_idx + 1}: 无有效候选样本")
            continue

        q_values_list = [s[4] for s in samples_raw]
        q_min = min(q_values_list)
        q_max = max(q_values_list)

        print(f"路径 {path_idx + 1}: 生成 {len(samples_raw)} 个候选样本")
        print(f"  Q值范围: [{q_min:.4f}, {q_max:.4f}]")

        samples_final = []
        for state, sim, len_diff, rob, q_value in samples_raw:
            if q_max - q_min > 1e-6:
                q_normalized = (q_value - q_min) / (q_max - q_min)
            else:
                q_normalized = 0.5

            q_complement = 1.0 - q_normalized
            score = 0.28 * sim + 0.1 * len_diff + 0.19 * rob + 0.43 * q_complement
            samples_final.append((state, score, sim, len_diff, rob, q_complement))

        if samples_final:
            samples_final.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples_final[:top_k], base_dir=base_dir)
            print(f"路径 {path_idx + 1}: 保存 {min(top_k, len(samples_final))} 个最终样本")


# === 训练函数 ===
def enhanced_standard_generate_and_train_for_similar_paths(agent, similar_group, path_documents, episodes=500,
                                                           batch_size=32, steps_per_test=5, replay_times=10,
                                                           is_isolated=False):
    trained_paths = set()
    update_target_every = 100
    global_steps = 0

    for episode in range(episodes):
        episode_reward = 0
        episode_similarities = []
        episode_td_errors = []

        for path_idx in similar_group:
            if path_idx in trained_paths:
                continue

            file_path = os.path.join(path_documents,
                                     f"enhanced_standard_path{path_idx + 1}{'_isolated' if is_isolated else ''}.txt")
            path_data = load_path_data(file_path)
            target_path = targetPaths[path_idx]

            BATCH_SIZE = 50
            N_SAMPLES = 200
            N_STEPS = 10
            REPLAY_TIMES = 3

            for batch_start in range(0, N_SAMPLES, BATCH_SIZE):
                batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)
                for test_data in range(batch_start, batch_end):
                    if test_data >= len(path_data):
                        break

                    step_count = 0
                    state = path_data[test_data]
                    prev_state = None
                    prev_triggered = None
                    prev_reward = None

                    for step in range(N_STEPS):
                        legal_actions = []
                        for a in range(agent.action_dim):
                            dx, dy, dz = agent.decode_action(a)
                            cand_next = (state[0] + dx, state[1] + dy, state[2] + dz)
                            if is_valid_state(cand_next):
                                legal_actions.append(a)
                        if not legal_actions:
                            break

                        if random.random() < agent.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            norm_state = normalize_state(state)
                            state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        dx, dy, dz = agent.decode_action(action)
                        next_state = clip_state((state[0] + dx, state[1] + dy, state[2] + dz))

                        triggered = execute_Tr(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                        done = (step_count == N_STEPS - 1)

                        td_error = agent.store_transition(state, action, reward, next_state, done)
                        enhanced_standard_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                      next_state[0], next_state[1], next_state[2])

                        episode_similarities.append(jaccard_similarity(triggered, target_path))
                        episode_td_errors.append(td_error)

                        if prev_reward is not None:
                            enhanced_standard_metrics.record_action_improvement(reward, prev_reward)

                        prev_state = state
                        prev_triggered = triggered
                        prev_reward = reward
                        state = next_state
                        step_count += 1
                        episode_reward += reward
                        global_steps += 1

                        if global_steps % update_target_every == 0:
                            agent.update_target_model()

                for _ in range(REPLAY_TIMES):
                    if len(agent.replay_buffer) >= batch_size:
                        agent.train(batch_size)

            trained_paths.add(path_idx)

        avg_similarity = np.mean(episode_similarities) if episode_similarities else 0
        avg_td_error = np.mean(episode_td_errors) if episode_td_errors else 0
        enhanced_standard_metrics.record_episode_metrics(episode, episode_reward, avg_similarity, avg_td_error,
                                                         agent.epsilon, "similar")

        if episode % 10 == 0:
            agent.update_target_model()

        if len(trained_paths) == len(similar_group):
            break

    return agent


def enhanced_standard_generate_and_train_for_isolated_paths(agent_similar, agent_isolated, similar_group,
                                                            isolated_group, path_documents, episodes=500, batch_size=32,
                                                            is_isolated=True):
    trained_paths = set()
    update_target_every = 100
    global_steps = 0

    stage1_samples_pool = {}

    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_similar.replay_buffer.get_high_reward_samples(target_path, num_samples=100)
        stage1_samples_pool[path_idx] = high_reward_samples

    for episode in range(episodes):
        episode_reward = 0
        episode_similarities = []
        episode_td_errors = []

        for path_idx in isolated_group:
            if path_idx in trained_paths:
                continue

            file_path = os.path.join(path_documents, f"enhanced_standard_path{path_idx + 1}_isolated.txt")
            stage2_path_data = load_path_data(file_path)
            stage1_samples = stage1_samples_pool.get(path_idx, [])
            target_path = targetPaths[path_idx]

            BATCH_SIZE = 50
            N_SAMPLES_STAGE2 = min(140, len(stage2_path_data))
            N_SAMPLES_STAGE1 = min(60, len(stage1_samples))
            N_STEPS = 10
            REPLAY_TIMES = 3

            for batch_start in range(0, N_SAMPLES_STAGE2, BATCH_SIZE):
                batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES_STAGE2)

                for test_data_idx in range(batch_start, batch_end):
                    if test_data_idx >= len(stage2_path_data):
                        break

                    step_count = 0
                    state = stage2_path_data[test_data_idx]
                    prev_state = None
                    prev_triggered = None
                    prev_reward = None

                    for step in range(N_STEPS):
                        legal_actions = []
                        for a in range(agent_isolated.action_dim):
                            dx, dy, dz = agent_isolated.decode_action(a)
                            cand_next = (state[0] + dx, state[1] + dy, state[2] + dz)
                            if is_valid_state(cand_next):
                                legal_actions.append(a)

                        if not legal_actions:
                            break

                        if random.random() < agent_isolated.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            norm_state = normalize_state(state)
                            state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent_isolated.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        dx, dy, dz = agent_isolated.decode_action(action)
                        next_state = clip_state((state[0] + dx, state[1] + dy, state[2] + dz))

                        triggered = execute_Tr(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)

                        if target_path.issubset(triggered):
                            reward += 2.0

                        done = (step_count == N_STEPS - 1)

                        td_error = agent_isolated.store_transition(state, action, reward, next_state, done)
                        enhanced_standard_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                      next_state[0], next_state[1], next_state[2])

                        episode_similarities.append(jaccard_similarity(triggered, target_path))
                        episode_td_errors.append(td_error)

                        if prev_reward is not None:
                            enhanced_standard_metrics.record_action_improvement(reward, prev_reward)

                        prev_state = state
                        prev_triggered = triggered
                        prev_reward = reward
                        state = next_state
                        step_count += 1
                        episode_reward += reward
                        global_steps += 1

                        if global_steps % update_target_every == 0:
                            agent_isolated.update_target_model()

            if stage1_samples:
                for sample_idx in range(N_SAMPLES_STAGE1):
                    if sample_idx >= len(stage1_samples):
                        break

                    stage1_state_tuple, _, _, _ = stage1_samples[sample_idx]
                    step_count = 0
                    state = stage1_state_tuple
                    prev_state = None
                    prev_triggered = None
                    prev_reward = None

                    for step in range(N_STEPS):
                        legal_actions = []
                        for a in range(agent_isolated.action_dim):
                            dx, dy, dz = agent_isolated.decode_action(a)
                            cand_next = (state[0] + dx, state[1] + dy, state[2] + dz)
                            if is_valid_state(cand_next):
                                legal_actions.append(a)

                        if not legal_actions:
                            break

                        if random.random() < agent_isolated.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            norm_state = normalize_state(state)
                            state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent_isolated.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        dx, dy, dz = agent_isolated.decode_action(action)
                        next_state = clip_state((state[0] + dx, state[1] + dy, state[2] + dz))

                        triggered = execute_Tr(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                        reward *= 0.8

                        done = (step_count == N_STEPS - 1)

                        td_error = agent_isolated.store_transition(state, action, reward, next_state, done)
                        enhanced_standard_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                      next_state[0], next_state[1], next_state[2])

                        episode_similarities.append(jaccard_similarity(triggered, target_path))
                        episode_td_errors.append(td_error)

                        if prev_reward is not None:
                            enhanced_standard_metrics.record_action_improvement(reward, prev_reward)

                        prev_state = state
                        prev_triggered = triggered
                        prev_reward = reward
                        state = next_state
                        step_count += 1
                        episode_reward += reward
                        global_steps += 1

                        if global_steps % update_target_every == 0:
                            agent_isolated.update_target_model()

            for replay_round in range(REPLAY_TIMES):
                if len(agent_isolated.replay_buffer) >= batch_size:
                    agent_isolated.train(batch_size)

            trained_paths.add(path_idx)

        avg_similarity = np.mean(episode_similarities) if episode_similarities else 0
        avg_td_error = np.mean(episode_td_errors) if episode_td_errors else 0
        enhanced_standard_metrics.record_episode_metrics(episode, episode_reward, avg_similarity, avg_td_error,
                                                         agent_isolated.epsilon, "isolated")

        if episode % 10 == 0:
            agent_isolated.update_target_model()

        if len(trained_paths) == len(isolated_group):
            break

    return agent_isolated


# === Excel导出 ===
def append_metrics_to_combined_excel(metrics_collector, agent_similar, agent_isolated, similar_group, isolated_group,
                                     targetPaths, filepath, run_number):
    """将指标和最终样本写入综合Excel"""

    # ===== 性能指标 =====
    training_time = metrics_collector.end_time - metrics_collector.start_time if metrics_collector.end_time else 0
    avg_memory = metrics_collector.total_memory_usage / metrics_collector.memory_check_count if metrics_collector.memory_check_count > 0 else 0
    avg_similarity = np.mean(
        metrics_collector.final_output_similarities) if metrics_collector.final_output_similarities else 0
    avg_td_error = np.mean(metrics_collector.td_errors) if metrics_collector.td_errors else 0
    action_improve_rate = np.mean(metrics_collector.action_improvements) if metrics_collector.action_improvements else 0

    performance_row = {
        'Run': run_number,
        'Average Similarity': f"{avg_similarity:.4f}",
        'TD Error': f"{avg_td_error:.4f}",
        'Action Improve Rate': f"{action_improve_rate:.4f}",
        'Memory(MB)': f"{avg_memory:.2f}"
    }

    # ===== 最终样本数据 =====
    sample_rows = []

    # 相似组
    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_similar.replay_buffer.get_high_reward_samples(target_path, num_samples=20)

        for state_tuple, reward, sim, triggered in high_reward_samples:
            sample_rows.append({
                'Run': run_number,
                'Group Type': 'Similar',
                'Path ID': path_idx + 1,
                'X': state_tuple[0],
                'Y': state_tuple[1],
                'Z': state_tuple[2],
                'Similarity': f"{sim:.4f}",
                'Reward': f"{reward:.2f}",
                'Triggered Count': len(triggered),
                'Target Count': len(target_path),
                'Triggered Rules': str(sorted(triggered)),
                'Target Rules': str(sorted(target_path))
            })
    # 孤立组
    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)

        for state_tuple, reward, sim, triggered in high_reward_samples:
            sample_rows.append({
                'Run': run_number,
                'Group Type': 'Isolated',
                'Path ID': path_idx + 1,
                'X': state_tuple[0],
                'Y': state_tuple[1],
                'Z': state_tuple[2],
                'Similarity': f"{sim:.4f}",
                'Reward': f"{reward:.2f}",
                'Triggered Count': len(triggered),
                'Target Count': len(target_path),
                'Triggered Rules': str(sorted(triggered)),
                'Target Rules': str(sorted(target_path))
            })

    # ===== 写入Excel =====
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    if os.path.exists(filepath):
        try:
            df_performance = pd.read_excel(filepath, sheet_name='Metric')
            df_performance = pd.concat([df_performance, pd.DataFrame([performance_row])], ignore_index=True)
        except:
            df_performance = pd.DataFrame([performance_row])

        try:
            df_samples = pd.read_excel(filepath, sheet_name='final samples')
            df_samples = pd.concat([df_samples, pd.DataFrame(sample_rows)], ignore_index=True)
        except:
            df_samples = pd.DataFrame(sample_rows)
    else:
        df_performance = pd.DataFrame([performance_row])
        df_samples = pd.DataFrame(sample_rows)

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df_performance.to_excel(writer, sheet_name='Metric', index=False)
        df_samples.to_excel(writer, sheet_name='final samples', index=False)

        workbook = writer.book

        ws_performance = writer.sheets['Metric']
        ws_performance.column_dimensions['A'].width = 15
        for col in ['B', 'C', 'D', 'E']:
            ws_performance.column_dimensions[col].width = 20

        header_font = Font(bold=True, size=11)
        for cell in ws_performance[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in ws_performance.iter_rows(min_row=2, max_row=ws_performance.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        ws_samples = writer.sheets['final samples']
        column_widths = {
            'A': 12, 'B': 15, 'C': 12, 'D': 10, 'E': 10, 'F': 10,
            'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 40, 'L': 40
        }
        for col, width in column_widths.items():
            ws_samples.column_dimensions[col].width = width

        for cell in ws_samples[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in ws_samples.iter_rows(min_row=2, max_row=ws_samples.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    print(f"第 {run_number} 轮结果已写入: {filepath}")


# === 单次实验 ===
def run_single_experiment(run_number, results_save_dir):
    """运行单次实验"""
    print(f"\n{'=' * 80}")
    print(f"开始第 {run_number} 轮实验")
    print(f"{'=' * 80}\n")

    enhanced_standard_metrics.reset()
    enhanced_standard_metrics.start_training()

    model_path_similar = os.path.join(results_save_dir, f"run{run_number}_similar.pth")
    model_path_isolated = os.path.join(results_save_dir, f"run{run_number}_isolated.pth")
    path_documents = r"D:\Experiment\CNN\DQNNEW\path_samples_enhanced_standard"

    similar_group, isolated_group = group_paths_by_similarity(targetPaths)

    # 第1轮生成样本
    if run_number == 1:
        generate_samples_for_similar_paths(similar_group, num_total=2000, top_k=200)

    replay_buffer = StandardExperienceReplay(capacity=10000)
    state_dim = 3
    action_dim = 30
    agent = StandardDQNAgent(state_dim, action_dim, replay_buffer)

    agent = enhanced_standard_generate_and_train_for_similar_paths(agent, similar_group, path_documents, episodes=500,
                                                                   batch_size=32, is_isolated=False)

    os.makedirs(os.path.dirname(model_path_similar), exist_ok=True)
    torch.save({
        'model_state_dict': agent.model.state_dict(),
        'optimizer_state_dict': agent.optimizer.state_dict(),
        'epsilon': agent.epsilon
    }, model_path_similar)

    # 第1轮生成孤立组样本
    if run_number == 1:
        generate_samples_for_isolated_paths_standard(agent, isolated_group, num_total=2000, top_k=200)

    isolated_replay_buffer = StandardExperienceReplay(capacity=15000)
    agent_isolated = StandardDQNAgent(state_dim, action_dim, isolated_replay_buffer)

    try:
        checkpoint = torch.load(model_path_similar)
        agent_isolated.model.load_state_dict(checkpoint['model_state_dict'])
        agent_isolated.target_model.load_state_dict(checkpoint['model_state_dict'])
        agent_isolated.epsilon = checkpoint.get('epsilon', 0.5)
    except Exception as e:
        pass

    agent_isolated = enhanced_standard_generate_and_train_for_isolated_paths(
        agent_similar=agent,
        agent_isolated=agent_isolated,
        similar_group=similar_group,
        isolated_group=isolated_group,
        path_documents=path_documents,
        episodes=500,
        batch_size=32,
        is_isolated=True
    )

    os.makedirs(os.path.dirname(model_path_isolated), exist_ok=True)
    torch.save({
        'model_state_dict': agent_isolated.model.state_dict(),
        'optimizer_state_dict': agent_isolated.optimizer.state_dict(),
        'epsilon': agent_isolated.epsilon
    }, model_path_isolated)

    enhanced_standard_metrics.end_training()

    # 记录最终样本
    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, _, sim, triggered in high_reward_samples:
            enhanced_standard_metrics.record_final_output_sample(triggered, target_path)

    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, _, sim, triggered in high_reward_samples:
            enhanced_standard_metrics.record_final_output_sample(triggered, target_path)

    # 写入Excel
    combined_excel_path = os.path.join(results_save_dir, "综合结果.xlsx")

    append_metrics_to_combined_excel(
        metrics_collector=enhanced_standard_metrics,
        agent_similar=agent,
        agent_isolated=agent_isolated,
        similar_group=similar_group,
        isolated_group=isolated_group,
        targetPaths=targetPaths,
        filepath=combined_excel_path,
        run_number=run_number
    )

    avg_similarity = np.mean(enhanced_standard_metrics.final_output_similarities)
    print(f"\n第 {run_number} 轮实验完成:")
    print(f"  平均相似度: {avg_similarity:.4f}")
    print(f"  总步数: {enhanced_standard_metrics.step_count}")


if __name__ == "__main__":
    results_save_dir = r"D:\Experiment\CNN\实验结果"
    os.makedirs(results_save_dir, exist_ok=True)

    NUM_RUNS = 20

    print("=" * 80)
    print(f"开始 {NUM_RUNS} 轮DQN实验(标准版无优先级)")
    print(f"范围: X[{X_MIN}, {X_MAX}], Y[{Y_MIN}, {Y_MAX}], Z[{Z_MIN}, {Z_MAX}]")
    print("=" * 80)

    for run in range(1, NUM_RUNS + 1):
        try:
            run_single_experiment(run, results_save_dir)
        except Exception as e:
            print(f"\n第 {run} 轮实验出错: {str(e)}")
            import traceback
            traceback.print_exc()
            continue

    print("\n" + "=" * 80)
    print(f"所有 {NUM_RUNS} 轮实验完成")
    print(f"结果保存在: {results_save_dir}")
    print("=" * 80)