import os
import random
from collections import deque
import numpy as np
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# --- 全局状态范围配置 ---
STATE_MIN_X, STATE_MAX_X = 2, 100
STATE_MIN_Y, STATE_MAX_Y = 20, 150
STATE_MIN_Z, STATE_MAX_Z = 30, 200

def generate_input():
    return [
        random.randint(STATE_MIN_X, STATE_MAX_X),
        random.randint(STATE_MIN_Y, STATE_MAX_Y),
        random.randint(STATE_MIN_Z, STATE_MAX_Z)
    ]

def execute_Tr(x, y, z):
    triggered = set()

    # --- 分支 1-15 (原 fault_y * fault_z / (fault_x + 1) > 85 的变异) ---
    if ((y * z) / (x + 1) > 85) != ((y * z) / (x + 2) > 85): triggered.add(1)
    if ((y * z) / (x + 1) > 85) != ((y * z) / (y + 1) > 85): triggered.add(2)
    if ((y * z) / (x + 1) > 85) != ((y * z) / (z + 1) > 85): triggered.add(3)
    if ((y * z) / (x + 1) > 85) != ((y * z) / (x * 1) > 85): triggered.add(4)
    if ((y * z) / (x + 1) > 85) != ((y * y) / (x + 1) > 85): triggered.add(5)
    if ((y * z) / (x + 1) > 85) != ((y * x) / (x + 1) > 85): triggered.add(6)
    if ((y * z) / (x + 1) > 85) != ((x * z) / (x + 1) > 85): triggered.add(7)
    if ((y * z) / (x + 1) > 85) != ((z * z) / (x + 1) > 85): triggered.add(8)
    if ((y * z) / (x + 1) > 85) != ((10 * z) / (x + 1) > 85): triggered.add(9)
    if ((y * z) / (x + 1) > 85) != ((y * z) - (x + 1) > 85): triggered.add(10)
    if ((y * z) / (x + 1) > 85) != ((y * z) / (x + 1) > 105): triggered.add(11)
    if ((y * z) / (x + 1) > 85) != ((y * z) / (x - 1) > 85): triggered.add(12)
    if ((y * z) / (x + 1) > 85) != ((y * 2 * z) / (x + 1) > 85): triggered.add(13)
    if ((y * z) / (x + 1) > 85) != ((y / 2 * z) / (x + 1) > 85): triggered.add(14)
    if ((y * z) / (x + 1) > 85) != ((y * 15) / (x + 1) > 85): triggered.add(15)

    # --- 分支 16-28 (原 (fault_z - fault_x) < 0.25 * fault_y 的变异) ---
    if ((z - x) < 0.25 * y) != ((y - x) < 0.25 * y): triggered.add(16)
    if ((z - x) < 0.25 * y) != ((z * 1.2 - x) < 0.25 * y): triggered.add(17)
    if ((z - x) < 0.25 * y) != ((z - x) < 0.3 * y): triggered.add(18)
    if ((z - x) < 0.25 * y) != ((z - x) < 0.4 * y): triggered.add(19)
    if ((z - x) < 0.25 * y) != ((z - x) < 0.25 * x): triggered.add(20)
    if ((z - x) < 0.25 * y) != ((z - x) < 0.25 * z): triggered.add(21)
    if ((z - x) < 0.25 * y) != ((y - x) < 0.25 * y): triggered.add(22)
    if ((z - x) < 0.25 * y) != ((z - x * 0.7) < 0.25 * y): triggered.add(23)
    if ((z - x) < 0.25 * y) != ((z - x) < 0.25 ** y): triggered.add(24)
    if ((z - x) < 0.25 * y) != ((z / x) < 0.25 * y): triggered.add(25)
    if ((z - x) < 0.25 * y) != ((z + x) < 0.25 * y): triggered.add(26)
    if ((z - x) < 0.25 * y) != ((80 - x) < 0.25 * y): triggered.add(27)
    if ((z - x) < 0.25 * y) != ((z - 60) < 0.25 * y): triggered.add(28)

    # --- 分支 29-44 (原 (fault_x^3 + fault_y^3) < fault_z^2 的变异) ---
    if ((x ** 3 + y ** 3) < z ** 2) != ((y ** 3 + y ** 3) < z ** 2): triggered.add(29)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x * 3 + y ** 3) < z ** 2): triggered.add(30)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 - y ** 3) < z ** 2): triggered.add(31)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + x ** 3) < z ** 2): triggered.add(32)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * 3) < z ** 2): triggered.add(33)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + z ** 3) < z ** 2): triggered.add(34)
    if ((x ** 3 + y ** 3) < z ** 2) != ((z ** 3 + y ** 3) < z ** 2): triggered.add(35)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < x ** 2): triggered.add(36)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < y ** 2): triggered.add(37)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z * 2): triggered.add(38)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 2.8) < z ** 2): triggered.add(39)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2.4 + y ** 3) < z ** 2): triggered.add(40)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 2.3): triggered.add(41)
    if ((x ** 3 + y ** 3) < z ** 2) != ((20 ** 3 + y ** 3) < z ** 2): triggered.add(42)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + 15 ** 3) < z ** 2): triggered.add(43)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < 50 ** 2): triggered.add(44)

    # --- 分支 45-56 (原 x/(y+0.01)>4.5 and y/(z+0.01)<0.22 的变异) ---
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((z / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22): triggered.add(45)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 1 and (y / (z + 0.01)) < 0.22): triggered.add(46)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 0.5 and (y / (z + 0.01)) < 0.22): triggered.add(47)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (x + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22): triggered.add(48)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x * 0.6 / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0): triggered.add(49)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 7.5 and (y / (z + 0.01)) < 0.22): triggered.add(50)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 4.5 and (z / (z + 0.01)) < 0.22): triggered.add(51)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 4.5 and (x / (z + 0.01)) < 0.22): triggered.add(52)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 4.5 and (y / (x + 0.01)) < 0.22): triggered.add(53)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 4.5 and (y / (y + 0.01)) < 0.22): triggered.add(54)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 2.5 and (y / (z + 0.01)) < 0.22): triggered.add(55)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.22): triggered.add(56)

    # --- 分支 57-68 (原 abs(x-y)>13 and abs(y-z)>17 and abs(x-z)<8 的变异) ---
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y - z) > 17 and abs(x * 2 - z) < 8): triggered.add(57)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x + y) > 13 and abs(y - z) > 17 and abs(x - z) < 8): triggered.add(58)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - z) > 13 and abs(y - z) > 17 and abs(x - z) < 8): triggered.add(59)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y * 1.1) > 13 and abs(y - z) > 17 and abs(x - z) < 8): triggered.add(60)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 18 and abs(y - z) > 17 and abs(x - z) < 8): triggered.add(61)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(x - z) > 17 and abs(x - z) < 8): triggered.add(62)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y - z * 0.9) > 17 and abs(x - z) < 8): triggered.add(63)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y * 1.4 - z) > 17 and abs(x - z) < 8): triggered.add(64)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y - z) > 20 and abs(x - z) < 8): triggered.add(65)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y - z) > 17 and abs(y - z) < 8): triggered.add(66)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y - z) > 17 and abs(x * 1.5 - z) < 8): triggered.add(67)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 4): triggered.add(68)

    # --- 分支 69-82 (原 (x>92 or x<6) and (y>87 or y<3) and (z>83 or z<2) 的变异) ---
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x * 3 > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(69)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x * x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(70)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x * y > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(71)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x * z > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(72)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 4) and (y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(73)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y * x > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(74)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y * y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(75)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y * z > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(76)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y > 87 or y < 3) and (z * x > 83 or z < 2)): triggered.add(77)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y > 87 or y < 3) and (z * y > 83 or z < 2)): triggered.add(78)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y > 87 or y < 3) and (z * z > 83 or z < 2)): triggered.add(79)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y > 87 or y < 3) and (z * 50 > 83 or z < 2)): triggered.add(80)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x * 60 > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(81)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y * 75 > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(82)

    return triggered


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0


targetPaths = [
    {2,7,8,10,13,17,23,24,26,28,57,59,60,61,62,63,64,66,67,68,74,75,76,82},  # A1
    {2,7,9,14,15,16,22,23,24,26,31,57,59,62,66,67,68,69,70,71,72,81},  # A2
    {2,7,8,10,13,16,22,25,27,29,30,42,45,46,47,55,56,57,74,75,76,82},  # A3
    {2,7,8,10,16,22,25,27,29,30,42,48,49,50,51,52,54,57,74,75,76,82},  # A4
    {2,7,9,11,14,15,23,24,26,31,57,59,62,65,66,67,69,70,71,72,81},  # A5
    {2,7,8,10,16,20,21,22,25,27,30,48,49,50,51,52,54,74,75,76,82},  # A6
    {1,2,3,6,7,8,9,11,14,15,16,22,24,26,27,31,57,59,62,66,67,68},  # A7
    {3,5,6,9,14,15,16,22,25,29,30,31,33,40,41,42,43,45,46,47},  # A8
    {3,5,6,9,14,15,16,22,29,30,31,33,39,40,41,42,43,45,46,47},  # A9
    {3,5,6,9,14,15,16,22,32,34,35,36,37,38,44,45,46,47},  # A10
    {3,4,5,6,10,12,13,16,22,26,27,31,57,59,62,66,67,68},  # A11
    {7,10,17,23,26,28,53,57,59,62,66,67,68,74,75,76,82},  # A12
    {7,8,10,16,18,19,20,21,22,25,27,57,59,62,66,67,68},  # A13
    {2,7,16,17,20,21,22,24,26,27,31,32,33,78,79,80},  # A14
    {2,6,7,9,18,19,25,28,31,32,33,77,78,79,80},  # A15
    {2,6,7,25,31,32,33,43,73},  # A16
    {2,7,8,9,11,14,15,26,31,58,60},  # A17
]


class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 6000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(targetPaths)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }


def compute_robustness(state, path):
    base = execute_Tr(*state)
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                        np.clip(state[0] + dx, STATE_MIN_X, STATE_MAX_X),
                        np.clip(state[1] + dy, STATE_MIN_Y, STATE_MAX_Y),
                        np.clip(state[2] + dz, STATE_MIN_Z, STATE_MAX_Z)
                    ])
                n_trig = execute_Tr(*neighbor)
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = targetPaths[target_path_idx]
    samples = []
    attempts = 0

    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(STATE_MIN_X, STATE_MAX_X),
            random.randint(STATE_MIN_Y, STATE_MAX_Y),
            random.randint(STATE_MIN_Z, STATE_MAX_Z)
        ])
        triggered = execute_Tr(*state)

        if not triggered:
            continue

        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)

        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }

        samples.append(sample_data)

    return samples


def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample[
                'robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)

        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]

    return selected_samples


def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = targetPaths[target_path_idx]
        samples = []
        attempts = 0

        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(STATE_MIN_X, STATE_MAX_X),
                random.randint(STATE_MIN_Y, STATE_MAX_Y),
                random.randint(STATE_MIN_Z, STATE_MAX_Z)
            ])
            triggered = execute_Tr(*state)

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)

            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)

        random.shuffle(samples)
        return samples[:config.top_k_samples]

    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")

        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)


def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}

    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates

    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}

        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config,
                                                         shared_candidates[path_idx])

            strategy_results[path_idx] = samples

        results[strategy_name] = strategy_results

    return results


def analyze_fitness_values(results, config):
    analysis_results = {}

    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []

        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])

        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }

        all_scores = []

        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])

        scores_array = np.array(all_scores)

        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)

        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })

        analysis_results[strategy_name] = analysis

    return analysis_results


def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []

    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)

    df = pd.DataFrame(df_data)
    return df, analysis_results


def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []

    print(f"Starting {num_runs} experiments...")

    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")

        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)

        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)

        print(f"Experiment {run_idx + 1} completed.")

    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df


def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')

    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width

    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []

    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]

        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)

    stats_df = pd.DataFrame(stats_data)

    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Results saved to: {output_path}")


def main():
    results_df = run_multiple_experiments(num_runs=20)

    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")

    save_results_to_excel(results_df, output_path)

    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)

    return results_df, output_path


if __name__ == "__main__":
    results, output_file = main()