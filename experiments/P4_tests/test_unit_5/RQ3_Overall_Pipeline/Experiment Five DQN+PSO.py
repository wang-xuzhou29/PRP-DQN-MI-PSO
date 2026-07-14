import os
import sys
import random
import time
from collections import deque
from datetime import datetime
from typing import Any, Dict, List, Optional, Set, Tuple

import numpy as np
import torch
import torch.nn as nn
import torch.optim as optim
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ========================================
# ========== experiment configuration parameters ==========
# ========================================
NUM_RUNS = 20

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === three-dimensional range settings ===
# Keep the current DQN state range used by the second script. To use a 0-500 range, modify this section only.
LIGHT_MIN = 2
LIGHT_MAX = 100
MOISTURE_MIN = 20
MOISTURE_MAX = 150
TEMP_MIN = 30
TEMP_MAX = 200
BOUNDS = {
    "light": (LIGHT_MIN, LIGHT_MAX),
    "moisture": (MOISTURE_MIN, MOISTURE_MAX),
    "temp": (TEMP_MIN, TEMP_MAX),
}

# === standard PSO parameters ===
PSO_W = 0.7
PSO_C1 = 1.5
PSO_C2 = 1.5
PSO_VMAX_RATIO = 0.2


def normalize_state(state):
    """Normalize the state to the [0, 1] range."""
    return np.array([
        (state[0] - BOUNDS["light"][0]) / (BOUNDS["light"][1] - BOUNDS["light"][0]),
        (state[1] - BOUNDS["moisture"][0]) / (BOUNDS["moisture"][1] - BOUNDS["moisture"][0]),
        (state[2] - BOUNDS["temp"][0]) / (BOUNDS["temp"][1] - BOUNDS["temp"][0]),
    ], dtype=np.float32)


def denormalize_state(normalized_state):
    """Restore the normalized state to the original range."""
    return np.array([
        normalized_state[0] * (BOUNDS["light"][1] - BOUNDS["light"][0]) + BOUNDS["light"][0],
        normalized_state[1] * (BOUNDS["moisture"][1] - BOUNDS["moisture"][0]) + BOUNDS["moisture"][0],
        normalized_state[2] * (BOUNDS["temp"][1] - BOUNDS["temp"][0]) + BOUNDS["temp"][0],
    ], dtype=float)


def generate_random_state():
    """Generate a random state within each dimension range."""
    light = np.random.randint(BOUNDS["light"][0], BOUNDS["light"][1] + 1)
    moisture = np.random.randint(BOUNDS["moisture"][0], BOUNDS["moisture"][1] + 1)
    temp = np.random.randint(BOUNDS["temp"][0], BOUNDS["temp"][1] + 1)
    return np.array([light, moisture, temp], dtype=float)


def clip_state(state):
    """Clip the state to each dimension boundary."""
    return np.array([
        np.clip(state[0], BOUNDS["light"][0], BOUNDS["light"][1]),
        np.clip(state[1], BOUNDS["moisture"][0], BOUNDS["moisture"][1]),
        np.clip(state[2], BOUNDS["temp"][0], BOUNDS["temp"][1]),
    ], dtype=float)


def is_state_valid(state):
    """Check whether the state is within all valid dimension ranges."""
    return (
        BOUNDS["light"][0] <= state[0] <= BOUNDS["light"][1]
        and BOUNDS["moisture"][0] <= state[1] <= BOUNDS["moisture"][1]
        and BOUNDS["temp"][0] <= state[2] <= BOUNDS["temp"][1]
    )


def execute_Tr(state_or_x, y=None, z=None):
    """
    执行目标函数，返回触发的分支集合。
    支持两种调用方式:
    - execute_Tr(position): position是3元素数组/元组
    - execute_Tr(x, y, z): 分别传入三个坐标值
    """
    if y is None and z is None:
        # 数组形式调用
        position = state_or_x
        x = int(np.clip(position[0], BOUNDS["light"][0], BOUNDS["light"][1]))
        y = int(np.clip(position[1], BOUNDS["moisture"][0], BOUNDS["moisture"][1]))
        z = int(np.clip(position[2], BOUNDS["temp"][0], BOUNDS["temp"][1]))
    else:
        # 分别传入x,y,z
        x = int(np.clip(state_or_x, BOUNDS["light"][0], BOUNDS["light"][1]))
        y = int(np.clip(y, BOUNDS["moisture"][0], BOUNDS["moisture"][1]))
        z = int(np.clip(z, BOUNDS["temp"][0], BOUNDS["temp"][1]))

    # 初始化分支覆盖数组 - 注意缩进，与if/else对齐
    triggered = set()

    # --- 分支 1-15 (原 fault_y * fault_z / (fault_x + 1) > 85 的变异) ---
    if ((y * z) / (x + 1) > 85) != ((y * z) / (x + 2) > 85): triggered.add(1)
    if ((y * z) / (x + 1) > 85) != ((y * z) / (y + 1) > 85): triggered.add(2)
    if ((y * z) / (x + 1) > 85) != ((y * z) / (z + 1) > 85): triggered.add(3)
    if ((y * z) / (x + 1) > 85) != ((y * z) / (x * 1) > 85): triggered.add(4)
    if ((y * z) / (x + 1) > 85) != ((y * y) / (x + 1) > 85): triggered.add(5)
    if ((y * z) / (x + 1) > 85) != ((y * x) / (x + 1) > 85): triggered.add(6)
    if ((y * z) / (x + 1) > 85) != ((x * z) / (x + 1) > 85): triggered.add(7)
    if ((y * z) / (x + 1) > 85) != ((z * z) / (x + 1) > 85): triggered.add(8)
    if ((y * z) / (x + 1) > 85) != ((10 * z) / (x + 1) > 85): triggered.add(9)
    if ((y * z) / (x + 1) > 85) != ((y * z) - (x + 1) > 85): triggered.add(10)
    if ((y * z) / (x + 1) > 85) != ((y * z) / (x + 1) > 105): triggered.add(11)
    if ((y * z) / (x + 1) > 85) != ((y * z) / (x - 1) > 85): triggered.add(12)
    if ((y * z) / (x + 1) > 85) != ((y * 2 * z) / (x + 1) > 85): triggered.add(13)
    if ((y * z) / (x + 1) > 85) != ((y / 2 * z) / (x + 1) > 85): triggered.add(14)
    if ((y * z) / (x + 1) > 85) != ((y * 15) / (x + 1) > 85): triggered.add(15)

    # --- 分支 16-28 (原 (fault_z - fault_x) < 0.25 * fault_y 的变异) ---
    if ((z - x) < 0.25 * y) != ((y - x) < 0.25 * y): triggered.add(16)
    if ((z - x) < 0.25 * y) != ((z * 1.2 - x) < 0.25 * y): triggered.add(17)
    if ((z - x) < 0.25 * y) != ((z - x) < 0.3 * y): triggered.add(18)
    if ((z - x) < 0.25 * y) != ((z - x) < 0.4 * y): triggered.add(19)
    if ((z - x) < 0.25 * y) != ((z - x) < 0.25 * x): triggered.add(20)
    if ((z - x) < 0.25 * y) != ((z - x) < 0.25 * z): triggered.add(21)
    if ((z - x) < 0.25 * y) != ((y - x) < 0.25 * y): triggered.add(22)
    if ((z - x) < 0.25 * y) != ((z - x * 0.7) < 0.25 * y): triggered.add(23)
    if ((z - x) < 0.25 * y) != ((z - x) < 0.25 ** y): triggered.add(24)
    if ((z - x) < 0.25 * y) != ((z / x) < 0.25 * y): triggered.add(25)
    if ((z - x) < 0.25 * y) != ((z + x) < 0.25 * y): triggered.add(26)
    if ((z - x) < 0.25 * y) != ((80 - x) < 0.25 * y): triggered.add(27)
    if ((z - x) < 0.25 * y) != ((z - 60) < 0.25 * y): triggered.add(28)

    # --- 分支 29-44 (原 (fault_x^3 + fault_y^3) < fault_z^2 的变异) ---
    if ((x ** 3 + y ** 3) < z ** 2) != ((y ** 3 + y ** 3) < z ** 2): triggered.add(29)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x * 3 + y ** 3) < z ** 2): triggered.add(30)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 - y ** 3) < z ** 2): triggered.add(31)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + x ** 3) < z ** 2): triggered.add(32)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * 3) < z ** 2): triggered.add(33)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + z ** 3) < z ** 2): triggered.add(34)
    if ((x ** 3 + y ** 3) < z ** 2) != ((z ** 3 + y ** 3) < z ** 2): triggered.add(35)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < x ** 2): triggered.add(36)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < y ** 2): triggered.add(37)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z * 2): triggered.add(38)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 2.8) < z ** 2): triggered.add(39)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2.4 + y ** 3) < z ** 2): triggered.add(40)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 2.3): triggered.add(41)
    if ((x ** 3 + y ** 3) < z ** 2) != ((20 ** 3 + y ** 3) < z ** 2): triggered.add(42)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + 15 ** 3) < z ** 2): triggered.add(43)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < 50 ** 2): triggered.add(44)

    # --- 分支 45-56 (原 x/(y+0.01)>4.5 and y/(z+0.01)<0.22 的变异) ---
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((z / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22): triggered.add(45)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 1 and (y / (z + 0.01)) < 0.22): triggered.add(46)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 0.5 and (y / (z + 0.01)) < 0.22): triggered.add(47)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (x + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22): triggered.add(48)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x * 0.6 / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0): triggered.add(49)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 7.5 and (y / (z + 0.01)) < 0.22): triggered.add(50)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 4.5 and (z / (z + 0.01)) < 0.22): triggered.add(51)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 4.5 and (x / (z + 0.01)) < 0.22): triggered.add(52)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 4.5 and (y / (x + 0.01)) < 0.22): triggered.add(53)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 4.5 and (y / (y + 0.01)) < 0.22): triggered.add(54)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 2.5 and (y / (z + 0.01)) < 0.22): triggered.add(55)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.22): triggered.add(56)

    # --- 分支 57-68 (原 abs(x-y)>13 and abs(y-z)>17 and abs(x-z)<8 的变异) ---
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y - z) > 17 and abs(x * 2 - z) < 8): triggered.add(57)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x + y) > 13 and abs(y - z) > 17 and abs(x - z) < 8): triggered.add(58)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - z) > 13 and abs(y - z) > 17 and abs(x - z) < 8): triggered.add(59)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y * 1.1) > 13 and abs(y - z) > 17 and abs(x - z) < 8): triggered.add(60)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 18 and abs(y - z) > 17 and abs(x - z) < 8): triggered.add(61)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(x - z) > 17 and abs(x - z) < 8): triggered.add(62)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y - z * 0.9) > 17 and abs(x - z) < 8): triggered.add(63)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y * 1.4 - z) > 17 and abs(x - z) < 8): triggered.add(64)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y - z) > 20 and abs(x - z) < 8): triggered.add(65)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y - z) > 17 and abs(y - z) < 8): triggered.add(66)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y - z) > 17 and abs(x * 1.5 - z) < 8): triggered.add(67)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 4): triggered.add(68)

    # --- 分支 69-82 (原 (x>92 or x<6) and (y>87 or y<3) and (z>83 or z<2) 的变异) ---
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x * 3 > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(69)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x * x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(70)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x * y > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(71)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x * z > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(72)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 4) and (y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(73)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y * x > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(74)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y * y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(75)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y * z > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(76)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y > 87 or y < 3) and (z * x > 83 or z < 2)): triggered.add(77)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y > 87 or y < 3) and (z * y > 83 or z < 2)): triggered.add(78)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y > 87 or y < 3) and (z * z > 83 or z < 2)): triggered.add(79)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y > 87 or y < 3) and (z * 50 > 83 or z < 2)): triggered.add(80)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x * 60 > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(81)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y * 75 > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(82)

    return triggered


target_paths = [
    {2,7,8,10,13,17,23,24,26,28,57,59,60,61,62,63,64,66,67,68,74,75,76,82},  # A1
    {2,7,9,14,15,16,22,23,24,26,31,57,59,62,66,67,68,69,70,71,72,81},  # A2
    {2,7,8,10,13,16,22,25,27,29,30,42,45,46,47,55,56,57,74,75,76,82},  # A3
    {2,7,8,10,16,22,25,27,29,30,42,48,49,50,51,52,54,57,74,75,76,82},  # A4
    {2,7,9,11,14,15,23,24,26,31,57,59,62,65,66,67,69,70,71,72,81},  # A5
    {2,7,8,10,16,20,21,22,25,27,30,48,49,50,51,52,54,74,75,76,82},  # A6
    {1,2,3,6,7,8,9,11,14,15,16,22,24,26,27,31,57,59,62,66,67,68},  # A7
    {3,5,6,9,14,15,16,22,25,29,30,31,33,40,41,42,43,45,46,47},  # A8
    {3,5,6,9,14,15,16,22,29,30,31,33,39,40,41,42,43,45,46,47},  # A9
    {3,5,6,9,14,15,16,22,32,34,35,36,37,38,44,45,46,47},  # A10
    {3,4,5,6,10,12,13,16,22,26,27,31,57,59,62,66,67,68},  # A11
    {7,10,17,23,26,28,53,57,59,62,66,67,68,74,75,76,82},  # A12
    {7,8,10,16,18,19,20,21,22,25,27,57,59,62,66,67,68},  # A13
    {2,7,16,17,20,21,22,24,26,27,31,32,33,78,79,80},  # A14
    {2,6,7,9,18,19,25,28,31,32,33,77,78,79,80},  # A15
    {2,6,7,25,31,32,33,43,73},  # A16
    {2,7,8,9,11,14,15,26,31,58,60},  # A17
]


def jaccard_similarity(set1: Set[int], set2: Set[int]) -> float:
    """If set1 covers target set2, the match score is treated as 1."""
    if set2.issubset(set1):
        return 1.0
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    return intersection / union if union != 0 else 0.0


def compute_reward(state, target_path, triggered):
    """Compute the DQN reward."""
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward


class SimpleReplayBuffer:
    def __init__(self, capacity=5000):
        self.buffer = deque(maxlen=capacity)

    def append(self, experience):
        self.buffer.append(experience)

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return random.sample(self.buffer, len(self.buffer))
        return random.sample(self.buffer, batch_size)

    def __len__(self):
        return len(self.buffer)


class SimpleDQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super().__init__()
        self.fc1 = nn.Linear(state_dim, 64)
        self.fc2 = nn.Linear(64, 32)
        self.fc3 = nn.Linear(32, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)


class SimpleDQNAgent:
    def __init__(self, state_dim, action_dim, learning_rate=0.001):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.epsilon = 0.3
        self.gamma = 0.99

        self.model = SimpleDQN(state_dim, action_dim).to(device)
        self.target_model = SimpleDQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())
        self.replay_buffer = SimpleReplayBuffer(capacity=5000)

    def decode_action(self, action_idx):
        """Decode the action."""
        light_deltas = [int((LIGHT_MAX - LIGHT_MIN) * p) for p in [0.05, 0.1, 0.2, 0.5, 0.7, -0.05, -0.1, -0.2, -0.5, -0.7]]
        moisture_deltas = [int((MOISTURE_MAX - MOISTURE_MIN) * p) for p in [0.05, 0.1, 0.2, 0.5, 0.7, -0.05, -0.1, -0.2, -0.5, -0.7]]
        temp_deltas = [int((TEMP_MAX - TEMP_MIN) * p) for p in [0.05, 0.1, 0.2, 0.5, 0.7, -0.05, -0.1, -0.2, -0.5, -0.7]]

        dim = action_idx // 10
        delta_idx = action_idx % 10

        if dim == 0:
            return (light_deltas[delta_idx], 0, 0)
        if dim == 1:
            return (0, moisture_deltas[delta_idx], 0)
        return (0, 0, temp_deltas[delta_idx])

    def store_transition(self, state, action, reward, next_state, done):
        self.replay_buffer.append((state, action, reward, next_state, done))

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return

        batch = self.replay_buffer.sample(batch_size)
        states, actions, rewards, next_states, dones = zip(*batch)

        states = torch.tensor(np.array(states), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array(next_states), dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))

        loss = nn.MSELoss()(current_q_values, target_q_values)

        self.optimizer.zero_grad()
        loss.backward()
        self.optimizer.step()

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())

    def get_best_samples(self, target_path, num_samples=20):
        """Screen DQN candidate samples from the replay buffer. This stage is counted separately in T_sample_screening."""
        if len(self.replay_buffer) == 0:
            return []

        samples_with_scores = []
        for state_norm, _, _, _, _ in self.replay_buffer.buffer:
            state_original = denormalize_state(state_norm)
            triggered = execute_Tr(state_original)
            reward = compute_reward(state_original, target_path, triggered)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_scores.append((state_original, reward, sim, triggered))

        samples_with_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_scores[:num_samples]


def train_dqn_for_path(path_idx, target_path, num_samples=200):
    """Train a standard DQN for a single path."""
    print(f"  Start training path {path_idx + 1} DQN model...")
    agent = SimpleDQNAgent(state_dim=3, action_dim=30)

    random_states = [normalize_state(generate_random_state()) for _ in range(num_samples)]

    STEPS_PER_SAMPLE = 3
    EPOCHS = 5
    BATCH_SIZE = 32

    step_count = 0
    for _ in range(EPOCHS):
        for state_norm in random_states:
            state_norm = tuple(state_norm)
            state_original = denormalize_state(state_norm)

            for step in range(STEPS_PER_SAMPLE):
                legal_actions = []
                for a in range(agent.action_dim):
                    dx, dy, dz = agent.decode_action(a)
                    next_state_candidate = (state_original[0] + dx, state_original[1] + dy, state_original[2] + dz)
                    if is_state_valid(next_state_candidate):
                        legal_actions.append(a)

                if not legal_actions:
                    break

                if random.random() < agent.epsilon:
                    action = random.choice(legal_actions)
                else:
                    state_tensor = torch.tensor(state_norm, dtype=torch.float32).unsqueeze(0).to(device)
                    with torch.no_grad():
                        q_values = agent.model(state_tensor)[0]
                    action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                dx, dy, dz = agent.decode_action(action)
                next_state_original = (state_original[0] + dx, state_original[1] + dy, state_original[2] + dz)
                next_state_norm = normalize_state(next_state_original)

                triggered = execute_Tr(next_state_original)
                reward = compute_reward(next_state_original, target_path, triggered)
                done = (step == STEPS_PER_SAMPLE - 1)

                agent.store_transition(state_norm, action, reward, next_state_norm, done)

                state_norm = next_state_norm
                state_original = next_state_original
                step_count += 1

                if step_count % 50 == 0 and len(agent.replay_buffer) >= BATCH_SIZE:
                    agent.train(BATCH_SIZE)

                if step_count % 100 == 0:
                    agent.update_target_model()

    print(f"  Path {path_idx + 1} training completed, replay buffer size: {len(agent.replay_buffer)}")
    return agent


class Particle:
    def __init__(self, initial_position=None):
        if initial_position is not None:
            self.position = clip_state(np.array(initial_position, dtype=float))
        else:
            self.position = np.array([
                np.random.uniform(BOUNDS["light"][0], BOUNDS["light"][1]),
                np.random.uniform(BOUNDS["moisture"][0], BOUNDS["moisture"][1]),
                np.random.uniform(BOUNDS["temp"][0], BOUNDS["temp"][1]),
            ], dtype=float)

        spans = np.array([
            BOUNDS["light"][1] - BOUNDS["light"][0],
            BOUNDS["moisture"][1] - BOUNDS["moisture"][0],
            BOUNDS["temp"][1] - BOUNDS["temp"][0],
        ], dtype=float)
        self.velocity = np.random.uniform(-0.1 * spans, 0.1 * spans)
        self.best_position = self.position.copy()
        self.best_fitness = -1.0
        self.fitness = -1.0


class BasicPSO:
    """
    Standard PSO: contains only standard velocity and position updates.
    Mutation coefficients, opposite particles, local-best reset, and other improved PSO mechanisms have been removed.
    """

    def __init__(self, target_path, swarm_size=20, dqn_samples=None):
        self.target_path = target_path
        self.swarm_size = swarm_size
        self.particles = []
        self.global_best_position = None
        self.global_best_fitness = -1.0

        if dqn_samples is not None and len(dqn_samples) > 0:
            num_direct = min(len(dqn_samples), swarm_size)
            for i in range(num_direct):
                state_tuple, _, _, _ = dqn_samples[i]
                self.particles.append(Particle(initial_position=state_tuple))

            while len(self.particles) < swarm_size:
                base_idx = len(self.particles) % len(dqn_samples)
                state_tuple, _, _, _ = dqn_samples[base_idx]
                perturbed = np.array(state_tuple, dtype=float) + np.random.randint(-10, 11, size=3)
                self.particles.append(Particle(initial_position=clip_state(perturbed)))
        else:
            self.particles = [Particle() for _ in range(swarm_size)]

        for particle in self.particles:
            particle.fitness = self.fitness_function(particle.position)
            particle.best_fitness = particle.fitness
            particle.best_position = particle.position.copy()

            if particle.fitness > self.global_best_fitness:
                self.global_best_fitness = particle.fitness
                self.global_best_position = particle.position.copy()

    def fitness_function(self, position):
        try:
            triggered = execute_Tr(position)
            return jaccard_similarity(triggered, self.target_path)
        except Exception:
            return 0.0

    def update(self):
        spans = np.array([
            BOUNDS["light"][1] - BOUNDS["light"][0],
            BOUNDS["moisture"][1] - BOUNDS["moisture"][0],
            BOUNDS["temp"][1] - BOUNDS["temp"][0],
        ], dtype=float)
        max_velocity = PSO_VMAX_RATIO * spans

        for particle in self.particles:
            r1 = np.random.random(3)
            r2 = np.random.random(3)

            particle.velocity = (
                PSO_W * particle.velocity
                + PSO_C1 * r1 * (particle.best_position - particle.position)
                + PSO_C2 * r2 * (self.global_best_position - particle.position)
            )
            particle.velocity = np.clip(particle.velocity, -max_velocity, max_velocity)

            particle.position = clip_state(particle.position + particle.velocity)
            particle.fitness = self.fitness_function(particle.position)

            if particle.fitness > particle.best_fitness:
                particle.best_fitness = particle.fitness
                particle.best_position = particle.position.copy()

            if particle.fitness > self.global_best_fitness:
                self.global_best_fitness = particle.fitness
                self.global_best_position = particle.position.copy()


def _set_header(ws, headers, widths, header_fill, header_font, border, center_align):
    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_row(ws, row_idx, row_data, border, alignment, fill=None):
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.border = border
        cell.alignment = alignment
        if fill is not None:
            cell.fill = fill


def export_time_metrics_to_excel(all_run_results, experiment_total_time=None, filename=None):
    """Export only timing metrics and per-path iteration counts."""
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"DQN_BasicPSO_Time_Metrics_{timestamp}.xlsx"

    output_dir = os.path.join(os.getcwd(), "results")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    num_runs = len(all_run_results)
    num_paths = len(target_paths)

    # Worksheet 1: runtime summary
    ws1 = wb.active
    ws1.title = "Runtime Summary"
    ws1.sheet_view.showGridLines = False
    headers1 = [
        "Run", "Number of Paths", "DQNTraining Total Time(s)", "Sample Screening Total Time(s)", "DQNDirect Check Total Time(s)",
        "PSOInitialization Total Time(s)", "PSOSearch Total Time(s)", "Total Algorithm Time(s)", "Total Iterations", "Average Iterations per Path"
    ]
    widths1 = [12, 10, 18, 18, 22, 18, 18, 16, 14, 18]
    _set_header(ws1, headers1, widths1, header_fill, header_font, border, center_align)

    for run_idx, run_result in enumerate(all_run_results, 1):
        path_results = run_result["paths"]
        total_dqn_train = sum(p["T_DQN_train"] for p in path_results)
        total_screen = sum(p["T_sample_screening"] for p in path_results)
        total_direct_check = sum(p["T_direct_check"] for p in path_results)
        total_pso_init = sum(p["T_PSO_init"] for p in path_results)
        total_pso_search = sum(p["T_PSO_search"] for p in path_results)
        total_iterations = sum(p["iterations"] for p in path_results)

        row_data = [
            f" {run_idx}",
            num_paths,
            round(total_dqn_train, 6),
            round(total_screen, 6),
            round(total_direct_check, 6),
            round(total_pso_init, 6),
            round(total_pso_search, 6),
            round(run_result["T_run_algorithm_total"], 6),
            int(total_iterations),
            round(total_iterations / num_paths, 2),
        ]
        _write_row(ws1, run_idx + 1, row_data, border, center_align, alternate_fill if run_idx % 2 == 0 else None)

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:J{num_runs + 1}"

    # Worksheet 2: path timing and iteration details
    ws2 = wb.create_sheet(title="Path Timing and Iteration Details")
    ws2.sheet_view.showGridLines = False
    headers2 = [
        "Run", "Path ID", "DQN training(s)", "Sample Screening Time(s)", "DQNDirect Check Time(s)",
        "PSOInitialization Time(s)", "PSOSearch Time(s)", "Path Total Algorithm Time(s)", "Iterations"
    ]
    widths2 = [12, 12, 18, 18, 22, 18, 18, 18, 12]
    _set_header(ws2, headers2, widths2, header_fill, header_font, border, center_align)

    row_idx = 2
    for run_idx, run_result in enumerate(all_run_results, 1):
        for path_result in run_result["paths"]:
            row_data = [
                f" {run_idx}",
                f"Path  {path_result['path_idx'] + 1}",
                round(path_result["T_DQN_train"], 6),
                round(path_result["T_sample_screening"], 6),
                round(path_result["T_direct_check"], 6),
                round(path_result["T_PSO_init"], 6),
                round(path_result["T_PSO_search"], 6),
                round(path_result["T_path_algorithm_total"], 6),
                int(path_result["iterations"]),
            ]
            _write_row(ws2, row_idx, row_data, border, center_align, alternate_fill if row_idx % 2 == 0 else None)
            row_idx += 1

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:I{row_idx - 1}"

    # Worksheet 3: path aggregate statistics
    ws3 = wb.create_sheet(title="Path Aggregate Statistics")
    ws3.sheet_view.showGridLines = False
    headers3 = [
        "Path ID", "Average DQN Training Time(s)", "Average Sample Screening Time(s)", "Average PSO Initialization Time(s)",
        "Average PSO Search Time(s)", "Average Path Total Algorithm Time(s)", "Average Iterations", "Minimum Iterations", "Maximum Iterations"
    ]
    widths3 = [12, 22, 22, 22, 22, 22, 16, 14, 14]
    _set_header(ws3, headers3, widths3, header_fill, header_font, border, center_align)

    for path_idx in range(num_paths):
        records = [run_result["paths"][path_idx] for run_result in all_run_results]
        iterations = [r["iterations"] for r in records]
        row_data = [
            f"Path  {path_idx + 1}",
            round(float(np.mean([r["T_DQN_train"] for r in records])), 6),
            round(float(np.mean([r["T_sample_screening"] for r in records])), 6),
            round(float(np.mean([r["T_PSO_init"] for r in records])), 6),
            round(float(np.mean([r["T_PSO_search"] for r in records])), 6),
            round(float(np.mean([r["T_path_algorithm_total"] for r in records])), 6),
            round(float(np.mean(iterations)), 2),
            int(np.min(iterations)),
            int(np.max(iterations)),
        ]
        _write_row(ws3, path_idx + 2, row_data, border, center_align, alternate_fill if (path_idx + 1) % 2 == 0 else None)

    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:I{num_paths + 1}"

    # Worksheet 4: experiment overview
    ws4 = wb.create_sheet(title="Experiment Overview")
    ws4.sheet_view.showGridLines = False
    headers4 = ["Metric", "Value"]
    widths4 = [32, 24]
    _set_header(ws4, headers4, widths4, header_fill, header_font, border, center_align)

    all_path_records = [p for run_result in all_run_results for p in run_result["paths"]]
    all_iterations = [p["iterations"] for p in all_path_records]
    overview_rows = [
        ["Experiment Runs", num_runs],
        ["Number of Target Paths", num_paths],
        ["Total Experiment Wall-clock Time(s)", round(experiment_total_time, 6) if experiment_total_time is not None else ""],
        ["Average Total Algorithm Time per Run(s)", round(float(np.mean([r["T_run_algorithm_total"] for r in all_run_results])), 6)],
        ["Average Total Algorithm Time per Path(s)", round(float(np.mean([p["T_path_algorithm_total"] for p in all_path_records])), 6)],
        ["Average DQN Training Time per Path(s)", round(float(np.mean([p["T_DQN_train"] for p in all_path_records])), 6)],
        ["Average Sample Screening Time per Path(s)", round(float(np.mean([p["T_sample_screening"] for p in all_path_records])), 6)],
        ["Average PSO Search Time per Path(s)", round(float(np.mean([p["T_PSO_search"] for p in all_path_records])), 6)],
        ["Average Iterations per Path", round(float(np.mean(all_iterations)), 2)],
        ["Maximum Iterations per Path", int(np.max(all_iterations))],
        ["Minimum Iterations per Path", int(np.min(all_iterations))],
    ]

    for idx, row_data in enumerate(overview_rows, 2):
        _write_row(ws4, idx, row_data, border, center_align, alternate_fill if idx % 2 == 0 else None)

    wb.save(filepath)
    print(f"\n Timing metrics exported to: {filepath}")
    return filepath


def run_single_experiment(run_num, max_iterations=3000):
    """Run one DQN + standard PSO ."""
    print(f"\n{'=' * 80}")
    print(f"Start run  {run_num}  run: DQN + standard PSO ")
    print(f"{'=' * 80}")

    run_start = time.perf_counter()
    path_results = []

    for path_idx, target_path in enumerate(target_paths):
        print(f"\nRun {run_num} run - Path {path_idx + 1}")
        path_total_start = time.perf_counter()

        # 1) DQN training
        dqn_train_start = time.perf_counter()
        agent = train_dqn_for_path(path_idx, target_path, num_samples=200)
        T_DQN_train = time.perf_counter() - dqn_train_start

        # 2) DQNSample Screening Time
        screening_start = time.perf_counter()
        dqn_samples = agent.get_best_samples(target_path, num_samples=20)
        T_sample_screening = time.perf_counter() - screening_start

        # 3) DQN
        direct_check_start = time.perf_counter()
        direct_solution_found = False
        for state_tuple, reward, sim, triggered in dqn_samples:
            if sim >= 1.0 or target_path.issubset(triggered):
                direct_solution_found = True
                break
        T_direct_check = time.perf_counter() - direct_check_start

        T_PSO_init = 0.0
        T_PSO_search = 0.0
        iterations_used = 0

        if direct_solution_found:
            print(
                f"  Path {path_idx + 1}: DQN | "
                f"DQN training {T_DQN_train:.6f}s | screening {T_sample_screening:.6f}s | iterations {iterations_used}"
            )
        else:
            # 4) standard PSOInitialization Time
            pso_init_start = time.perf_counter()
            pso = BasicPSO(target_path, swarm_size=20, dqn_samples=dqn_samples)
            T_PSO_init = time.perf_counter() - pso_init_start

            # 5) standard PSOSearch Time
            pso_search_start = time.perf_counter()
            iterations_used = max_iterations
            for iteration in range(max_iterations):
                pso.update()
                if pso.global_best_fitness >= 1.0:
                    iterations_used = iteration + 1
                    break
            T_PSO_search = time.perf_counter() - pso_search_start

            print(
                f"  Path {path_idx + 1}: "
                f"DQN training {T_DQN_train:.6f}s | screening {T_sample_screening:.6f}s | "
                f"PSO {T_PSO_init:.6f}s | PSO {T_PSO_search:.6f}s | "
                f"iterations {iterations_used}"
            )

        T_path_algorithm_total = time.perf_counter() - path_total_start

        path_results.append({
            "path_idx": path_idx,
            "T_DQN_train": T_DQN_train,
            "T_sample_screening": T_sample_screening,
            "T_direct_check": T_direct_check,
            "T_PSO_init": T_PSO_init,
            "T_PSO_search": T_PSO_search,
            "T_path_algorithm_total": T_path_algorithm_total,
            "iterations": iterations_used,
        })

    T_run_algorithm_total = time.perf_counter() - run_start
    print(f"\nRun {run_num} runcompleted | Total Algorithm Time {T_run_algorithm_total:.6f}s")
    return {
        "run_num": run_num,
        "paths": path_results,
        "T_run_algorithm_total": T_run_algorithm_total,
    }


def run_multiple_experiments(num_runs):
    print("\n" + "=" * 80)
    print(f"DQN + standard PSO - {num_runs}: Metric")
    print("=" * 80)
    print(f"Number of Target Paths: {len(target_paths)}")
    print("Statistics: DQN training, Sample Screening Time, PSOInitialization Time, PSOSearch Time, Path Iterations")
    print("=" * 80)

    all_run_results = []
    experiment_start = time.perf_counter()

    for run_num in range(1, num_runs + 1):
        run_result = run_single_experiment(run_num)
        all_run_results.append(run_result)

    experiment_total_time = time.perf_counter() - experiment_start
    print(f"\n{'=' * 80}")
    print(f"All {num_runs} runcompleted | Total Experiment Wall-clock Time {experiment_total_time:.6f}s")
    print(f"{'=' * 80}\n")
    return all_run_results, experiment_total_time


if __name__ == "__main__":
    print("=" * 80)
    print("DQN + standard PSO - MetricPath Iterations")
    print("=" * 80)
    print(f"Current configuration: Run = {NUM_RUNS}")
    print(f"Number of Paths: {len(target_paths)}")
    print(f"Device: {device}")
    print("=" * 80)

    if len(sys.argv) > 1:
        try:
            NUM_RUNS = int(sys.argv[1])
            print(f"Read from command line: Run = {NUM_RUNS}")
        except ValueError:
            print(f"Invalid command-line argument, using default number of runs {NUM_RUNS}")

    all_results, total_time = run_multiple_experiments(num_runs=NUM_RUNS)
    export_time_metrics_to_excel(all_results, total_time)
    print("\nProgram completed")
