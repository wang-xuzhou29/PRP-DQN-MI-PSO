import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """安全除法"""
    if denominator == 0:
        return default
    return numerator / denominator


def execute_Tr(x, y, z):
    """执行105分支逻辑，返回触发的分支集合"""
    # 初始化分支覆盖数组
    triggered = set()

    # --- 分支 1-15 (原 fault_y * fault_z / (fault_x + 1) > 85 的变异) ---
    if ((y * z) / (x + 1) > 85) != ((y * z) / (x + 2) > 85): triggered.add(1)
    if ((y * z) / (x + 1) > 85) != ((y * z) / (y + 1) > 85): triggered.add(2)
    if ((y * z) / (x + 1) > 85) != ((y * z) / (z + 1) > 85): triggered.add(3)
    if ((y * z) / (x + 1) > 85) != ((y * z) / (x * 1) > 85): triggered.add(4)
    if ((y * z) / (x + 1) > 85) != ((y * y) / (x + 1) > 85): triggered.add(5)
    if ((y * z) / (x + 1) > 85) != ((y * x) / (x + 1) > 85): triggered.add(6)
    if ((y * z) / (x + 1) > 85) != ((x * z) / (x + 1) > 85): triggered.add(7)
    if ((y * z) / (x + 1) > 85) != ((z * z) / (x + 1) > 85): triggered.add(8)
    if ((y * z) / (x + 1) > 85) != ((10 * z) / (x + 1) > 85): triggered.add(9)
    if ((y * z) / (x + 1) > 85) != ((y * z) - (x + 1) > 85): triggered.add(10)
    if ((y * z) / (x + 1) > 85) != ((y * z) / (x + 1) > 105): triggered.add(11)
    if ((y * z) / (x + 1) > 85) != ((y * z) / (x - 1) > 85): triggered.add(12)
    if ((y * z) / (x + 1) > 85) != ((y * 2 * z) / (x + 1) > 85): triggered.add(13)
    if ((y * z) / (x + 1) > 85) != ((y / 2 * z) / (x + 1) > 85): triggered.add(14)
    if ((y * z) / (x + 1) > 85) != ((y * 15) / (x + 1) > 85): triggered.add(15)

    # --- 分支 16-28 (原 (fault_z - fault_x) < 0.25 * fault_y 的变异) ---
    if ((z - x) < 0.25 * y) != ((y - x) < 0.25 * y): triggered.add(16)
    if ((z - x) < 0.25 * y) != ((z * 1.2 - x) < 0.25 * y): triggered.add(17)
    if ((z - x) < 0.25 * y) != ((z - x) < 0.3 * y): triggered.add(18)
    if ((z - x) < 0.25 * y) != ((z - x) < 0.4 * y): triggered.add(19)
    if ((z - x) < 0.25 * y) != ((z - x) < 0.25 * x): triggered.add(20)
    if ((z - x) < 0.25 * y) != ((z - x) < 0.25 * z): triggered.add(21)
    if ((z - x) < 0.25 * y) != ((y - x) < 0.25 * y): triggered.add(22)
    if ((z - x) < 0.25 * y) != ((z - x * 0.7) < 0.25 * y): triggered.add(23)
    if ((z - x) < 0.25 * y) != ((z - x) < 0.25 ** y): triggered.add(24)
    if ((z - x) < 0.25 * y) != ((z / x) < 0.25 * y): triggered.add(25)
    if ((z - x) < 0.25 * y) != ((z + x) < 0.25 * y): triggered.add(26)
    if ((z - x) < 0.25 * y) != ((80 - x) < 0.25 * y): triggered.add(27)
    if ((z - x) < 0.25 * y) != ((z - 60) < 0.25 * y): triggered.add(28)

    # --- 分支 29-44 (原 (fault_x^3 + fault_y^3) < fault_z^2 的变异) ---
    if ((x ** 3 + y ** 3) < z ** 2) != ((y ** 3 + y ** 3) < z ** 2): triggered.add(29)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x * 3 + y ** 3) < z ** 2): triggered.add(30)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 - y ** 3) < z ** 2): triggered.add(31)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + x ** 3) < z ** 2): triggered.add(32)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * 3) < z ** 2): triggered.add(33)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + z ** 3) < z ** 2): triggered.add(34)
    if ((x ** 3 + y ** 3) < z ** 2) != ((z ** 3 + y ** 3) < z ** 2): triggered.add(35)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < x ** 2): triggered.add(36)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < y ** 2): triggered.add(37)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z * 2): triggered.add(38)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 2.8) < z ** 2): triggered.add(39)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2.4 + y ** 3) < z ** 2): triggered.add(40)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 2.3): triggered.add(41)
    if ((x ** 3 + y ** 3) < z ** 2) != ((20 ** 3 + y ** 3) < z ** 2): triggered.add(42)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + 15 ** 3) < z ** 2): triggered.add(43)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < 50 ** 2): triggered.add(44)

    # --- 分支 45-56 (原 x/(y+0.01)>4.5 and y/(z+0.01)<0.22 的变异) ---
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((z / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22): triggered.add(45)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 1 and (y / (z + 0.01)) < 0.22): triggered.add(46)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 0.5 and (y / (z + 0.01)) < 0.22): triggered.add(47)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (x + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22): triggered.add(48)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x * 0.6 / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0): triggered.add(49)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 7.5 and (y / (z + 0.01)) < 0.22): triggered.add(50)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 4.5 and (z / (z + 0.01)) < 0.22): triggered.add(51)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 4.5 and (x / (z + 0.01)) < 0.22): triggered.add(52)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 4.5 and (y / (x + 0.01)) < 0.22): triggered.add(53)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 4.5 and (y / (y + 0.01)) < 0.22): triggered.add(54)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 2.5 and (y / (z + 0.01)) < 0.22): triggered.add(55)
    if ((x / (y + 0.01)) > 4.5 and (y / (z + 0.01)) < 0.22) != ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.22): triggered.add(56)

    # --- 分支 57-68 (原 abs(x-y)>13 and abs(y-z)>17 and abs(x-z)<8 的变异) ---
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y - z) > 17 and abs(x * 2 - z) < 8): triggered.add(57)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x + y) > 13 and abs(y - z) > 17 and abs(x - z) < 8): triggered.add(58)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - z) > 13 and abs(y - z) > 17 and abs(x - z) < 8): triggered.add(59)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y * 1.1) > 13 and abs(y - z) > 17 and abs(x - z) < 8): triggered.add(60)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 18 and abs(y - z) > 17 and abs(x - z) < 8): triggered.add(61)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(x - z) > 17 and abs(x - z) < 8): triggered.add(62)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y - z * 0.9) > 17 and abs(x - z) < 8): triggered.add(63)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y * 1.4 - z) > 17 and abs(x - z) < 8): triggered.add(64)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y - z) > 20 and abs(x - z) < 8): triggered.add(65)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y - z) > 17 and abs(y - z) < 8): triggered.add(66)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y - z) > 17 and abs(x * 1.5 - z) < 8): triggered.add(67)
    if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 4): triggered.add(68)

    # --- 分支 69-82 (原 (x>92 or x<6) and (y>87 or y<3) and (z>83 or z<2) 的变异) ---
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x * 3 > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(69)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x * x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(70)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x * y > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(71)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x * z > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(72)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 4) and (y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(73)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y * x > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(74)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y * y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(75)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y * z > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(76)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y > 87 or y < 3) and (z * x > 83 or z < 2)): triggered.add(77)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y > 87 or y < 3) and (z * y > 83 or z < 2)): triggered.add(78)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y > 87 or y < 3) and (z * z > 83 or z < 2)): triggered.add(79)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y > 87 or y < 3) and (z * 50 > 83 or z < 2)): triggered.add(80)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x * 60 > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(81)
    if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y * 75 > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(82)

    return triggered


def execute_validation_rules(dx: int, dy: int, dz: int) -> Set[int]:
    """执行验证规则，返回触发的分支集合（调用105分支版本）"""
    return execute_Tr(dx, dy, dz)


def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """计算适应度"""
    generated_path = execute_validation_rules(int(particle[0]), int(particle[1]), int(particle[2]))

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """基本PSO优化器"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # 范围: x(2,100), y(20,150), z(30,200)
        self.bounds = bounds if bounds else [(2, 100), (20, 150), (30, 200)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """初始化粒子"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """更新速度和位置"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """对目标路径进行PSO优化"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(
                            int(particles[i][0]), int(particles[i][1]), int(particles[i][2])
                        ),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(
                int(gbest_particle[0]), int(gbest_particle[1]), int(gbest_particle[2])
            ),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """对所有路径运行PSO"""

    print(f"\n{'=' * 70}")
    print(f"Baseline PSO - 路径优化")
    print(f"{'=' * 70}")
    print(f"粒子数: {n_particles}, 最大迭代次数: {max_iterations}")
    print(f"路径数量: {len(target_paths)}")
    print(f"范围: x(2,100), y(20,150), z(30,200)")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"路径 {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "完美求解" if result['success'] else f"(适应度: {result['best_fitness']:.3f})"
        print(f"{status} | 耗时: {result['time']:.2f}s | 迭代次数: {result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f"求解成功: {success_count}/{len(target_paths)} ({success_rate:.1f}%) | 总耗时: {total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """多次运行实验"""

    print(f"\n{'=' * 70}")
    print(f"Baseline PSO - {num_runs}次运行")
    print(f"{'=' * 70}")
    print(f"粒子数: {n_particles}, 最大迭代次数: {max_iterations}, 路径数: {len(target_paths)}")
    print(f"范围: x(2,100), y(20,150), z(30,200)")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"--- 第 {run_idx}/{num_runs} 次运行 ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f"求解成功: {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs}次运行完成 | 总耗时: {total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """导出Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # Sheet 1: 运行汇总
    ws1 = wb.active
    ws1.title = "运行汇总"
    ws1.sheet_view.showGridLines = False

    headers = ["运行", "成功率", "求解/总数", "平均适应度", "平均迭代次数", "耗时(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f"运行 {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # Sheet 2: 路径统计
    ws2 = wb.create_sheet(title="路径统计")
    ws2.sheet_view.showGridLines = False

    headers2 = ["路径ID", "求解/总数", "成功率", "平均适应度", "平均迭代次数", "最小迭代次数", "最大迭代次数"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"路径 {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # Sheet 3: 详细结果
    ws3 = wb.create_sheet(title="详细结果")
    ws3.sheet_view.showGridLines = False

    headers3 = ["路径ID", "运行", "位置(x,y,z)", "适应度", "迭代次数", "触发路径"]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"Run {run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # Sheet 4: 目标路径
    ws4 = wb.create_sheet(title="目标路径")
    ws4.sheet_view.showGridLines = False

    headers4 = ["路径ID", "目标路径", "分支数"]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"路径 {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f"Excel已保存: {filename}")
    print(f"{'=' * 70}")
    print(f"包含工作表:")
    print(f"  1. 运行汇总      - {len(all_results)}次运行汇总")
    print(f"  2. 路径统计      - 每条路径的统计")
    print(f"  3. 详细结果      - 每次运行每条路径的详细结果")
    print(f"  4. 目标路径      - 目标路径定义")
    print(f"{'=' * 70}\n")

    return filename


def main():
    # 目标路径定义（105分支版本）
    target_paths = [
        {2,7,8,10,13,17,23,24,26,28,57,59,60,61,62,63,64,66,67,68,74,75,76,82},  # A1
        {2,7,9,14,15,16,22,23,24,26,31,57,59,62,66,67,68,69,70,71,72,81},  # A2
        {2,7,8,10,13,16,22,25,27,29,30,42,45,46,47,55,56,57,74,75,76,82},  # A3
        {2,7,8,10,16,22,25,27,29,30,42,48,49,50,51,52,54,57,74,75,76,82},  # A4
        {2,7,9,11,14,15,23,24,26,31,57,59,62,65,66,67,69,70,71,72,81},  # A5
        {2,7,8,10,16,20,21,22,25,27,30,48,49,50,51,52,54,74,75,76,82},  # A6
        {1,2,3,6,7,8,9,11,14,15,16,22,24,26,27,31,57,59,62,66,67,68},  # A7
        {3,5,6,9,14,15,16,22,25,29,30,31,33,40,41,42,43,45,46,47},  # A8
        {3,5,6,9,14,15,16,22,29,30,31,33,39,40,41,42,43,45,46,47},  # A9
        {3,5,6,9,14,15,16,22,32,34,35,36,37,38,44,45,46,47},  # A10
        {3,4,5,6,10,12,13,16,22,26,27,31,57,59,62,66,67,68},  # A11
        {7,10,17,23,26,28,53,57,59,62,66,67,68,74,75,76,82},  # A12
        {7,8,10,16,18,19,20,21,22,25,27,57,59,62,66,67,68},  # A13
        {2,7,16,17,20,21,22,24,26,27,31,32,33,78,79,80},  # A14
        {2,6,7,9,18,19,25,28,31,32,33,77,78,79,80},  # A15
        {2,6,7,25,31,32,33,43,73},  # A16
        {2,7,8,9,11,14,15,26,31,58,60},  # A17
    ]

    print("=" * 70)
    print("Baseline PSO - 105分支版本")
    print("=" * 70)
    print(f"范围: x(2,100), y(1,105), z(1,110)")
    print(f"路径数: {len(target_paths)}")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("程序完成")


if __name__ == "__main__":
    main()