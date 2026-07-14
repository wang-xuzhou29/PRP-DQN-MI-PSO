import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import json
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment

# === 设备设置 ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === 状态空间范围 ===
dx_min, dx_max = 2, 100
dy_min, dy_max = 20, 150
dz_min, dz_max = 30, 200


def normalize_state(state):
    dx, dy, dz = state
    normalized_dx = (dx - dx_min) / (dx_max - dx_min)
    normalized_dy = (dy - dy_min) / (dy_max - dy_min)
    normalized_dz = (dz - dz_min) / (dz_max - dz_min)
    return (normalized_dx, normalized_dy, normalized_dz)


def denormalize_state(normalized_state):
    norm_dx, norm_dy, norm_dz = normalized_state
    dx = int(norm_dx * (dx_max - dx_min) + dx_min)
    dy = int(norm_dy * (dy_max - dy_min) + dy_min)
    dz = int(norm_dz * (dz_max - dz_min) + dz_min)
    return (dx, dy, dz)


def is_valid_state(state):
    dx, dy, dz = state
    return (dx_min <= dx <= dx_max and
            dy_min <= dy <= dy_max and
            dz_min <= dz <= dz_max)


def clip_state(state):
    dx, dy, dz = state
    return (
        max(dx_min, min(dx_max, dx)),
        max(dy_min, min(dy_max, dy)),
        max(dz_min, min(dz_max, dz))
    )


class PrioritizedMetricsCollector:
    def __init__(self, experiment_name="Prioritized_DQN"):
        self.experiment_name = experiment_name
        self.start_time = None
        self.end_time = None
        self.total_reward = 0
        self.td_errors = []
        self.final_output_similarities = []
        self.action_improvements = []
        self.total_memory_usage = 0
        self.memory_check_count = 0
        self.step_count = 0
        self.episode_rewards = []
        self.episode_similarities = []
        self.episode_td_errors = []
        self.episode_epsilon_values = []
        self.episode_memory_usage = []
        self.priority_statistics = []
        self.importance_weights = []
        self.high_priority_samples_ratio = []
        self.priority_distribution_stats = []
        self.similar_paths_performance = []
        self.isolated_paths_performance = []
        self.milestone_data = {}
        self.convergence_window = 20
        self.convergence_threshold = 0.02
        self.convergence_detected_episode = None
        self.sample_efficiency_data = []
        self.performance_milestones = [0.6, 0.7, 0.75, 0.8]
        self.learning_curve_characteristics = {}
        self.early_vs_late_performance = {}

    def reset(self):
        self.start_time = None
        self.end_time = None
        self.total_reward = 0
        self.td_errors = []
        self.final_output_similarities = []
        self.action_improvements = []
        self.total_memory_usage = 0
        self.memory_check_count = 0
        self.step_count = 0
        self.episode_rewards = []
        self.episode_similarities = []
        self.episode_td_errors = []
        self.episode_epsilon_values = []
        self.episode_memory_usage = []
        self.priority_statistics = []
        self.importance_weights = []
        self.high_priority_samples_ratio = []
        self.priority_distribution_stats = []
        self.similar_paths_performance = []
        self.isolated_paths_performance = []
        self.milestone_data = {}
        self.convergence_detected_episode = None
        self.sample_efficiency_data = []
        self.learning_curve_characteristics = {}
        self.early_vs_late_performance = {}

    def start_training(self):
        self.start_time = time.time()

    def end_training(self):
        self.end_time = time.time()

    def record_step_metrics(self, reward, td_error, triggered, target_path, priority=None, is_weight=None):
        self.step_count += 1
        self.total_reward += reward
        self.td_errors.append(td_error)
        if priority is not None:
            self.priority_statistics.append(priority)
        if is_weight is not None:
            self.importance_weights.append(is_weight)
        try:
            process = psutil.Process(os.getpid())
            current_memory = process.memory_info().rss / 1024 / 1024
            self.total_memory_usage += current_memory
            self.memory_check_count += 1
        except:
            pass

    def record_episode_metrics(self, episode, episode_reward, avg_similarity, avg_td_error, epsilon,
                               path_group="similar", priority_stats=None):
        self.episode_rewards.append(episode_reward)
        self.episode_similarities.append(avg_similarity)
        self.episode_td_errors.append(avg_td_error)
        self.episode_epsilon_values.append(epsilon)
        try:
            process = psutil.Process(os.getpid())
            current_memory = process.memory_info().rss / 1024 / 1024
            self.episode_memory_usage.append(current_memory)
        except:
            self.episode_memory_usage.append(0)
        if priority_stats:
            self.priority_distribution_stats.append({
                'episode': episode,
                'mean_priority': priority_stats.get('mean_priority', 0),
                'max_priority': priority_stats.get('max_priority', 0),
                'min_priority': priority_stats.get('min_priority', 0),
                'high_priority_ratio': priority_stats.get('high_priority_ratio', 0)
            })
        if path_group == "similar":
            self.similar_paths_performance.append({
                'episode': episode, 'reward': episode_reward,
                'similarity': avg_similarity, 'td_error': avg_td_error
            })
        else:
            self.isolated_paths_performance.append({
                'episode': episode, 'reward': episode_reward,
                'similarity': avg_similarity, 'td_error': avg_td_error
            })
        if episode in [50, 100, 150, 200, 250, 300, 400, 450, 500]:
            current_memory = self.episode_memory_usage[-1] if self.episode_memory_usage else 0
            self.milestone_data[episode] = {
                'avg_reward': np.mean(self.episode_rewards[-10:]) if len(self.episode_rewards) >= 10 else episode_reward,
                'avg_similarity': avg_similarity,
                'avg_td_error': avg_td_error,
                'epsilon': epsilon,
                'memory_usage': current_memory,
                'total_steps': self.step_count,
                'priority_stats': priority_stats
            }
        self._check_convergence(episode)
        self._check_performance_milestones(episode, avg_similarity)

    def _check_convergence(self, episode):
        if len(self.episode_similarities) >= self.convergence_window and self.convergence_detected_episode is None:
            recent_similarities = self.episode_similarities[-self.convergence_window:]
            if np.std(recent_similarities) < self.convergence_threshold:
                self.convergence_detected_episode = episode

    def _check_performance_milestones(self, episode, similarity):
        for milestone in self.performance_milestones:
            if similarity >= milestone and not any(data[1] == milestone for data in self.sample_efficiency_data):
                self.sample_efficiency_data.append((episode, milestone, self.step_count))

    def record_final_output_sample(self, triggered, target_path):
        if len(triggered | target_path) > 0:
            similarity = len(triggered & target_path) / len(triggered | target_path)
        else:
            similarity = 0.0
        self.final_output_similarities.append(similarity)

    def record_action_improvement(self, current_reward, prev_reward):
        if prev_reward is not None:
            improvement = current_reward - prev_reward
            self.action_improvements.append(1 if improvement > 0 else 0)


prioritized_metrics = PrioritizedMetricsCollector("Prioritized_DQN_Enhanced")


def safe_divide(a, b, default=0.0):
    if b == 0 or abs(b) < 1e-10:
        return default
    return a / b


def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward


def execute_Tr(x, y, z):
    triggered = set()
    safe_x = x if x != 0 else 1e-10
    safe_y = y if y != 0 else 1e-10
    safe_z = z if z != 0 else 1e-10

    try:
        denom1 = x + 1
        val1 = safe_divide(y * z, denom1)
        if (val1 > 85) != (safe_divide(y * z, x + 2) > 85): triggered.add(1)
        if (val1 > 85) != (safe_divide(y * z, y + 1) > 85): triggered.add(2)
        if (val1 > 85) != (safe_divide(y * z, z + 1) > 85): triggered.add(3)
        if (val1 > 85) != (safe_divide(y * z, safe_x) > 85): triggered.add(4)
        if (val1 > 85) != (safe_divide(y * y, denom1) > 85): triggered.add(5)
        if (val1 > 85) != (safe_divide(y * x, denom1) > 85): triggered.add(6)
        if (val1 > 85) != (safe_divide(x * z, denom1) > 85): triggered.add(7)
        if (val1 > 85) != (safe_divide(z * z, denom1) > 85): triggered.add(8)
        if (val1 > 85) != (safe_divide(10 * z, denom1) > 85): triggered.add(9)
        if (val1 > 85) != ((y * z) - (x + 1) > 85): triggered.add(10)
        if (val1 > 85) != (val1 > 105): triggered.add(11)
        if (val1 > 85) != (safe_divide(y * z, x - 1 if x != 1 else 1e-10) > 85): triggered.add(12)
        if (val1 > 85) != (safe_divide(y * 2 * z, denom1) > 85): triggered.add(13)
        if (val1 > 85) != (safe_divide(y / 2 * z, denom1) > 85): triggered.add(14)
        if (val1 > 85) != (safe_divide(y * 15, denom1) > 85): triggered.add(15)
    except:
        pass

    try:
        if ((z - x) < 0.25 * y) != ((y - x) < 0.25 * y): triggered.add(16)
        if ((z - x) < 0.25 * y) != ((z * 1.2 - x) < 0.25 * y): triggered.add(17)
        if ((z - x) < 0.25 * y) != ((z - x) < 0.3 * y): triggered.add(18)
        if ((z - x) < 0.25 * y) != ((z - x) < 0.4 * y): triggered.add(19)
        if ((z - x) < 0.25 * y) != ((z - x) < 0.25 * x): triggered.add(20)
        if ((z - x) < 0.25 * y) != ((z - x) < 0.25 * z): triggered.add(21)
        if ((z - x) < 0.25 * y) != ((y - x) < 0.25 * y): triggered.add(22)
        if ((z - x) < 0.25 * y) != ((z - x * 0.7) < 0.25 * y): triggered.add(23)
        if ((z - x) < 0.25 * y) != ((z - x) < 0.25 ** y): triggered.add(24)
        if ((z - x) < 0.25 * y) != (safe_divide(z, safe_x) < 0.25 * y): triggered.add(25)
        if ((z - x) < 0.25 * y) != ((z + x) < 0.25 * y): triggered.add(26)
        if ((z - x) < 0.25 * y) != ((80 - x) < 0.25 * y): triggered.add(27)
        if ((z - x) < 0.25 * y) != ((z - 60) < 0.25 * y): triggered.add(28)
    except:
        pass

    try:
        if ((x ** 3 + y ** 3) < z ** 2) != ((y ** 3 + y ** 3) < z ** 2): triggered.add(29)
        if ((x ** 3 + y ** 3) < z ** 2) != ((x * 3 + y ** 3) < z ** 2): triggered.add(30)
        if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 - y ** 3) < z ** 2): triggered.add(31)
        if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + x ** 3) < z ** 2): triggered.add(32)
        if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * 3) < z ** 2): triggered.add(33)
        if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + z ** 3) < z ** 2): triggered.add(34)
        if ((x ** 3 + y ** 3) < z ** 2) != ((z ** 3 + y ** 3) < z ** 2): triggered.add(35)
        if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < x ** 2): triggered.add(36)
        if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < y ** 2): triggered.add(37)
        if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z * 2): triggered.add(38)
        if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 2.8) < z ** 2): triggered.add(39)
        if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2.4 + y ** 3) < z ** 2): triggered.add(40)
        if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 2.3): triggered.add(41)
        if ((x ** 3 + y ** 3) < z ** 2) != ((20 ** 3 + y ** 3) < z ** 2): triggered.add(42)
        if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + 15 ** 3) < z ** 2): triggered.add(43)
        if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < 50 ** 2): triggered.add(44)
    except:
        pass

    try:
        denom_y = y + 0.01
        denom_z = z + 0.01
        val45_1 = safe_divide(x, denom_y)
        val45_2 = safe_divide(y, denom_z)
        if (val45_1 > 4.5 and val45_2 < 0.22) != (safe_divide(z, denom_y) > 4.5 and val45_2 < 0.22): triggered.add(45)
        if (val45_1 > 4.5 and val45_2 < 0.22) != (val45_1 > 1 and val45_2 < 0.22): triggered.add(46)
        if (val45_1 > 4.5 and val45_2 < 0.22) != (val45_1 > 0.5 and val45_2 < 0.22): triggered.add(47)
        if (val45_1 > 4.5 and val45_2 < 0.22) != (safe_divide(x, x + 0.01) > 4.5 and val45_2 < 0.22): triggered.add(48)
        if (val45_1 > 4.5 and val45_2 < 0.22) != (safe_divide(x * 0.6, denom_y) > 4.5 and val45_2 < 0): triggered.add(49)
        if (val45_1 > 4.5 and val45_2 < 0.22) != (val45_1 > 7.5 and val45_2 < 0.22): triggered.add(50)
        if (val45_1 > 4.5 and val45_2 < 0.22) != (val45_1 > 4.5 and safe_divide(z, denom_z) < 0.22): triggered.add(51)
        if (val45_1 > 4.5 and val45_2 < 0.22) != (val45_1 > 4.5 and safe_divide(x, denom_z) < 0.22): triggered.add(52)
        if (val45_1 > 4.5 and val45_2 < 0.22) != (val45_1 > 4.5 and safe_divide(y, x + 0.01) < 0.22): triggered.add(53)
        if (val45_1 > 4.5 and val45_2 < 0.22) != (val45_1 > 4.5 and safe_divide(y, denom_y) < 0.22): triggered.add(54)
        if (val45_1 > 4.5 and val45_2 < 0.22) != (val45_1 > 2.5 and val45_2 < 0.22): triggered.add(55)
        if (val45_1 > 4.5 and val45_2 < 0.22) != (val45_1 > 3.5 and val45_2 < 0.22): triggered.add(56)
    except:
        pass

    try:
        if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y - z) > 17 and abs(x * 2 - z) < 8): triggered.add(57)
        if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x + y) > 13 and abs(y - z) > 17 and abs(x - z) < 8): triggered.add(58)
        if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - z) > 13 and abs(y - z) > 17 and abs(x - z) < 8): triggered.add(59)
        if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y * 1.1) > 13 and abs(y - z) > 17 and abs(x - z) < 8): triggered.add(60)
        if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 18 and abs(y - z) > 17 and abs(x - z) < 8): triggered.add(61)
        if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(x - z) > 17 and abs(x - z) < 8): triggered.add(62)
        if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y - z * 0.9) > 17 and abs(x - z) < 8): triggered.add(63)
        if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y * 1.4 - z) > 17 and abs(x - z) < 8): triggered.add(64)
        if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y - z) > 20 and abs(x - z) < 8): triggered.add(65)
        if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y - z) > 17 and abs(y - z) < 8): triggered.add(66)
        if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y - z) > 17 and abs(x * 1.5 - z) < 8): triggered.add(67)
        if (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 8) != (abs(x - y) > 13 and abs(y - z) > 17 and abs(x - z) < 4): triggered.add(68)
    except:
        pass

    try:
        if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x * 3 > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(69)
        if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x * x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(70)
        if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x * y > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(71)
        if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x * z > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(72)
        if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 4) and (y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(73)
        if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y * x > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(74)
        if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y * y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(75)
        if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y * z > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(76)
        if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y > 87 or y < 3) and (z * x > 83 or z < 2)): triggered.add(77)
        if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y > 87 or y < 3) and (z * y > 83 or z < 2)): triggered.add(78)
        if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y > 87 or y < 3) and (z * z > 83 or z < 2)): triggered.add(79)
        if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y > 87 or y < 3) and (z * 50 > 83 or z < 2)): triggered.add(80)
        if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x * 60 > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(81)
        if ((x > 92 or x < 6) and (y > 87 or y < 3) and (z > 83 or z < 2)) != ((x > 92 or x < 6) and (y * 75 > 87 or y < 3) and (z > 83 or z < 2)): triggered.add(82)
    except:
        pass

    return triggered


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0


def compute_path_similarity_matrix(paths):
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix


targetPaths = [
    {2,7,8,10,13,17,23,24,26,28,57,59,60,61,62,63,64,66,67,68,74,75,76,82},
    {2,7,9,14,15,16,22,23,24,26,31,57,59,62,66,67,68,69,70,71,72,81},
    {2,7,8,10,13,16,22,25,27,29,30,42,45,46,47,55,56,57,74,75,76,82},
    {2,7,8,10,16,22,25,27,29,30,42,48,49,50,51,52,54,57,74,75,76,82},
    {2,7,9,11,14,15,23,24,26,31,57,59,62,65,66,67,69,70,71,72,81},
    {2,7,8,10,16,20,21,22,25,27,30,48,49,50,51,52,54,74,75,76,82},
    {1,2,3,6,7,8,9,11,14,15,16,22,24,26,27,31,57,59,62,66,67,68},
    {3,5,6,9,14,15,16,22,25,29,30,31,33,40,41,42,43,45,46,47},
    {3,5,6,9,14,15,16,22,29,30,31,33,39,40,41,42,43,45,46,47},
    {3,5,6,9,14,15,16,22,32,34,35,36,37,38,44,45,46,47},
    {3,4,5,6,10,12,13,16,22,26,27,31,57,59,62,66,67,68},
    {7,10,17,23,26,28,53,57,59,62,66,67,68,74,75,76,82},
    {7,8,10,16,18,19,20,21,22,25,27,57,59,62,66,67,68},
    {2,7,16,17,20,21,22,24,26,27,31,32,33,78,79,80},
    {2,6,7,9,18,19,25,28,31,32,33,77,78,79,80},
    {2,6,7,25,31,32,33,43,73},
    {2,7,8,9,11,14,15,26,31,58,60},
]


def group_paths_by_similarity(paths):
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.mean(avg_sim_scores)
    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]
    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)
    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group


class PrioritizedExperienceReplay:
    def __init__(self, capacity=10000, alpha=0.6, beta_start=0.4, beta_frames=100000):
        self.capacity = capacity
        self.alpha = alpha
        self.beta_start = beta_start
        self.beta_frames = beta_frames
        self.frame = 1
        self.buffer = []
        self.priorities = np.zeros(capacity, dtype=np.float32)
        self.pos = 0
        self.size = 0
        self.max_priority = 1.0
        self.min_priority = 1.0

    def beta(self):
        return min(1.0, self.beta_start + (1.0 - self.beta_start) * self.frame / self.beta_frames)

    def append(self, experience):
        if len(self.buffer) < self.capacity:
            self.buffer.append(experience)
        else:
            self.buffer[self.pos] = experience
        self.priorities[self.pos] = self.max_priority
        self.pos = (self.pos + 1) % self.capacity
        self.size = min(self.size + 1, self.capacity)

    def sample(self, batch_size):
        if self.size < batch_size:
            return [], [], []
        priorities = self.priorities[:self.size]
        probs = priorities ** self.alpha
        probs /= probs.sum()
        indices = np.random.choice(self.size, batch_size, p=probs, replace=False)
        unique_batch = []
        unique_indices = []
        seen_states = set()
        for idx in indices:
            experience = self.buffer[idx]
            state_tensor = experience[0]
            state_tuple = tuple(state_tensor.cpu().numpy().flatten())
            if state_tuple not in seen_states:
                seen_states.add(state_tuple)
                unique_batch.append(experience)
                unique_indices.append(idx)
        if len(unique_batch) < batch_size:
            remaining_indices = [i for i in range(self.size) if i not in unique_indices]
            if remaining_indices:
                remaining_probs = priorities[remaining_indices] ** self.alpha
                remaining_probs /= remaining_probs.sum()
                needed = batch_size - len(unique_batch)
                additional_indices = np.random.choice(remaining_indices, min(needed, len(remaining_indices)), p=remaining_probs, replace=False)
                for idx in additional_indices:
                    experience = self.buffer[idx]
                    state_tensor = experience[0]
                    state_tuple = tuple(state_tensor.cpu().numpy().flatten())
                    if state_tuple not in seen_states:
                        seen_states.add(state_tuple)
                        unique_batch.append(experience)
                        unique_indices.append(idx)
        total = len(self.buffer)
        unique_indices = np.array(unique_indices)
        weights = (total * probs[unique_indices]) ** (-self.beta())
        weights /= weights.max()
        self.frame += 1
        return unique_batch, unique_indices, weights

    def update_priorities(self, indices, priorities):
        for idx, priority in zip(indices, priorities):
            if idx < self.size:
                self.priorities[idx] = priority
                self.max_priority = max(self.max_priority, priority)
                self.min_priority = min(self.min_priority, priority)

    def get_priority_statistics(self):
        if self.size == 0:
            return None
        priorities = self.priorities[:self.size]
        return {
            'mean_priority': np.mean(priorities),
            'max_priority': np.max(priorities),
            'min_priority': np.min(priorities),
            'high_priority_ratio': np.mean(priorities > np.mean(priorities))
        }

    def __len__(self):
        return self.size

    def get_high_reward_samples(self, target_path, num_samples=20):
        if len(self.buffer) == 0:
            return []
        samples_with_recalculated_scores = []
        for experience in self.buffer:
            state_tensor = experience[0]
            state_tuple = tuple(state_tensor.cpu().numpy().flatten().astype(int))
            triggered = execute_Tr(*state_tuple)
            new_reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_recalculated_scores.append((state_tuple, new_reward, sim, triggered))
        samples_with_recalculated_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_recalculated_scores[:num_samples]


def load_path_data(file_path):
    path_data = []
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
            for line in lines[2:]:
                parts = line.strip().split("\t")
                state = tuple(map(int, parts[0].split()))
                path_data.append(state)
    except FileNotFoundError:
        print(f"警告: 文件未找到 - {file_path}")
    except Exception as e:
        print(f"警告: 读取文件出错 - {file_path}: {e}")
    return path_data


class DQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, 128)
        self.fc2 = nn.Linear(128, 64)
        self.fc3 = nn.Linear(64, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)


class PrioritizedDQNAgent:
    def __init__(self, state_dim, action_dim, replay_buffer, gamma=0.99, epsilon=1.0,
                 epsilon_decay=0.995, epsilon_min=0.1, learning_rate=0.001):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer
        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        delta_values = [35, 25, 10, 5, 2, -2, -5, -10, -25, -35]
        dim = action_idx // 10
        delta_idx = action_idx % 10
        if dim == 0:
            return (delta_values[delta_idx], 0, 0)
        elif dim == 1:
            return (0, delta_values[delta_idx], 0)
        elif dim == 2:
            return (0, 0, delta_values[delta_idx])
        return (0, 0, 0)

    def act(self, state):
        if random.random() < self.epsilon:
            return random.randrange(self.action_dim)
        normalized_state = normalize_state(state)
        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)
        return torch.argmax(q_values, dim=1).item()

    def store_transition(self, state, action, reward, next_state, done):
        normalized_state = normalize_state(state)
        normalized_next_state = normalize_state(next_state)
        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        next_state_tensor = torch.tensor(normalized_next_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)
            next_q_values = self.target_model(next_state_tensor)
            max_next_q_values = next_q_values.max(1)[0]
            target_q_values = reward + (self.gamma * max_next_q_values * (1 - done))
            td_error = torch.abs(q_values[0][action] - target_q_values).item()
        self.replay_buffer.append((state_tensor, action, reward, next_state_tensor, done, td_error))
        return td_error

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return
        batch, batch_indices, is_weights = self.replay_buffer.sample(batch_size)
        if not batch:
            return
        states, actions, rewards, next_states, dones, _ = zip(*batch)
        states = torch.tensor(np.array([s.cpu().numpy().flatten() for s in states]), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array([ns.cpu().numpy().flatten() for ns in next_states]), dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)
        is_weights = torch.tensor(is_weights, dtype=torch.float32).to(device)
        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))
        td_errors = torch.abs(current_q_values - target_q_values)
        loss = (is_weights * (current_q_values - target_q_values) ** 2).mean()
        self.optimizer.zero_grad()
        loss.backward()
        self.optimizer.step()
        new_priorities = td_errors.detach().cpu().numpy() + 1e-6
        self.replay_buffer.update_priorities(batch_indices, new_priorities)
        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())


def generate_samples_for_similar_paths(similar_group_indices, num_total=2000, top_k=200):
    def jaccard_similarity_local(a, b):
        if not a and not b:
            return 1.0
        return len(a & b) / len(a | b) if a | b else 0.0

    def compute_robustness(state, path):
        base = execute_Tr(*state)
        if not base:
            return 0.0
        rob, neighbors = 0.0, 0
        for ddx in [-1, 0, 1]:
            for ddy in [-1, 0, 1]:
                for ddz in [-1, 0, 1]:
                    if ddx == ddy == ddz == 0:
                        continue
                    neighbor_state = (state[0] + ddx, state[1] + ddy, state[2] + ddz)
                    if not is_valid_state(neighbor_state):
                        continue
                    neighbor = clip_state(neighbor_state)
                    n_trig = execute_Tr(*neighbor)
                    if not n_trig:
                        continue
                    rob += jaccard_similarity_local(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"prioritized_path{path_id}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Prioritized Path {path_id}\n")
            f.write("dx dy dz\tScore\tSimilarity\tLengthDiff\tRobustness\n")
            for s in samples:
                dx, dy, dz = s[0]
                f.write(f"{dx} {dy} {dz}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_prioritized"
    for path_idx in similar_group_indices:
        path = targetPaths[path_idx]
        samples = []
        attempts = 0
        while len(samples) < top_k and attempts < num_total * 5:
            attempts += 1
            state = (np.random.randint(dx_min, dx_max + 1), np.random.randint(dy_min, dy_max + 1), np.random.randint(dz_min, dz_max + 1))
            triggered = execute_Tr(*state)
            if not triggered:
                continue
            sim = jaccard_similarity_local(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            score = 0.55 * sim + 0.25 * len_diff + 0.2 * rob
            samples.append((state, score, sim, len_diff, rob))
        if samples:
            samples.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples[:top_k], base_dir=base_dir)


def prioritized_generate_and_train_for_similar_paths(agent, similar_group, path_documents, episodes=500,
                                                     batch_size=32, steps_per_test=3, replay_times=1,
                                                     is_isolated=False):
    trained_paths = set()
    global_replay_count = 0
    for episode in range(episodes):
        episode_reward = 0
        episode_similarities = []
        episode_td_errors = []
        for path_idx in similar_group:
            if path_idx in trained_paths:
                continue
            file_path = os.path.join(path_documents, f"prioritized_path{path_idx + 1}{'_isolated' if is_isolated else ''}.txt")
            path_data = load_path_data(file_path)
            if not path_data:
                trained_paths.add(path_idx)
                continue
            target_path = targetPaths[path_idx]
            BATCH_SIZE = 50
            N_SAMPLES = min(200, len(path_data))
            N_STEPS = 3
            N_BATCHES = 4
            PATH_REPEAT = 5
            for repeat in range(PATH_REPEAT):
                for batch_idx in range(N_BATCHES):
                    batch_start = batch_idx * BATCH_SIZE
                    batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)
                    for test_data_idx in range(batch_start, batch_end):
                        if test_data_idx >= len(path_data):
                            break
                        state = path_data[test_data_idx]
                        prev_state = None
                        prev_triggered = None
                        prev_reward = None
                        for step in range(N_STEPS):
                            legal_actions = []
                            for a in range(agent.action_dim):
                                ddx, ddy, ddz = agent.decode_action(a)
                                cand_next = (state[0] + ddx, state[1] + ddy, state[2] + ddz)
                                if is_valid_state(cand_next):
                                    legal_actions.append(a)
                            if not legal_actions:
                                break
                            if random.random() < agent.epsilon:
                                action = random.choice(legal_actions)
                            else:
                                normalized_state = normalize_state(state)
                                state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
                                with torch.no_grad():
                                    q_values = agent.model(state_tensor)[0]
                                action = legal_actions[torch.argmax(q_values[legal_actions]).item()]
                            ddx, ddy, ddz = agent.decode_action(action)
                            next_state = clip_state((state[0] + ddx, state[1] + ddy, state[2] + ddz))
                            triggered = execute_Tr(*next_state)
                            reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                            done = (step == N_STEPS - 1)
                            td_error = agent.store_transition(state, action, reward, next_state, done)
                            priority_stats = agent.replay_buffer.get_priority_statistics()
                            current_priority = priority_stats['mean_priority'] if priority_stats else 0
                            prioritized_metrics.record_step_metrics(reward, td_error, triggered, target_path, current_priority, None)
                            episode_similarities.append(jaccard_similarity(triggered, target_path))
                            episode_td_errors.append(td_error)
                            if prev_reward is not None:
                                prioritized_metrics.record_action_improvement(reward, prev_reward)
                            prev_state = state
                            prev_triggered = triggered
                            prev_reward = reward
                            state = next_state
                            episode_reward += reward
                    if len(agent.replay_buffer) >= batch_size:
                        agent.train(batch_size)
                        global_replay_count += 1
                        if global_replay_count % 2 == 0:
                            agent.update_target_model()
            trained_paths.add(path_idx)
        avg_similarity = np.mean(episode_similarities) if episode_similarities else 0
        avg_td_error = np.mean(episode_td_errors) if episode_td_errors else 0
        priority_stats = agent.replay_buffer.get_priority_statistics()
        prioritized_metrics.record_episode_metrics(episode, episode_reward, avg_similarity, avg_td_error, agent.epsilon, "similar", priority_stats)
        if len(trained_paths) == len(similar_group):
            break
    return agent


def generate_samples_for_isolated_paths_prioritized(agent_similar, isolated_group_indices, num_total=2000, top_k=200):
    def compute_q_value_normalized_complement(state, agent):
        normalized_state = normalize_state(state)
        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = agent.model(state_tensor)
        all_q_values = q_values[0].cpu().numpy()
        q_min = all_q_values.min()
        q_max = all_q_values.max()
        if q_max - q_min > 1e-6:
            normalized_q = (all_q_values.max() - q_min) / (q_max - q_min)
        else:
            normalized_q = 0.0
        return 1.0 - normalized_q

    def compute_robustness(state, path):
        base = execute_Tr(*state)
        if not base:
            return 0.0
        rob, neighbors = 0.0, 0
        for ddx in [-1, 0, 1]:
            for ddy in [-1, 0, 1]:
                for ddz in [-1, 0, 1]:
                    if ddx == ddy == ddz == 0:
                        continue
                    neighbor_state = (state[0] + ddx, state[1] + ddy, state[2] + ddz)
                    if not is_valid_state(neighbor_state):
                        continue
                    neighbor = clip_state(neighbor_state)
                    n_trig = execute_Tr(*neighbor)
                    if not n_trig:
                        continue
                    rob += jaccard_similarity(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"prioritized_path{path_id}_isolated.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Prioritized Isolated Path {path_id}\n")
            f.write("dx dy dz\tScore\tSimilarity\tLengthDiff\tRobustness\tQ_value_complement\n")
            for s in samples:
                dx, dy, dz = s[0]
                f.write(f"{dx} {dy} {dz}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\t{s[5]:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_prioritized"
    for path_idx in isolated_group_indices:
        path = targetPaths[path_idx]
        samples = []
        attempts = 0
        while len(samples) < top_k and attempts < num_total * 5:
            attempts += 1
            state = (np.random.randint(dx_min, dx_max + 1), np.random.randint(dy_min, dy_max + 1), np.random.randint(dz_min, dz_max + 1))
            triggered = execute_Tr(*state)
            if not triggered:
                continue
            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            q_value_complement = compute_q_value_normalized_complement(state, agent_similar)
            score = 0.28 * sim + 0.1 * len_diff + 0.19 * rob + 0.43 * q_value_complement
            samples.append((state, score, sim, len_diff, rob, q_value_complement))
        if samples:
            samples.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples[:top_k], base_dir=base_dir)


def prioritized_generate_and_train_for_isolated_paths(agent_similar, agent_isolated, similar_group,
                                                      isolated_group, path_documents, episodes=500, batch_size=32,
                                                      is_isolated=True):
    trained_paths = set()
    global_replay_count = 0
    stage1_samples_pool = {}
    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_similar.replay_buffer.get_high_reward_samples(target_path, num_samples=100)
        stage1_samples_pool[path_idx] = high_reward_samples
    for episode in range(episodes):
        episode_reward = 0
        episode_similarities = []
        episode_td_errors = []
        for path_idx in isolated_group:
            if path_idx in trained_paths:
                continue
            file_path = os.path.join(path_documents, f"prioritized_path{path_idx + 1}_isolated.txt")
            stage2_path_data = load_path_data(file_path)
            if not stage2_path_data:
                trained_paths.add(path_idx)
                continue
            stage1_samples = stage1_samples_pool.get(path_idx, [])
            target_path = targetPaths[path_idx]
            BATCH_SIZE = 50
            N_SAMPLES_STAGE2 = min(140, len(stage2_path_data))
            N_SAMPLES_STAGE1 = min(60, len(stage1_samples))
            N_STEPS = 3
            N_BATCHES_STAGE2 = (N_SAMPLES_STAGE2 + BATCH_SIZE - 1) // BATCH_SIZE
            N_BATCHES_STAGE1 = (N_SAMPLES_STAGE1 + BATCH_SIZE - 1) // BATCH_SIZE
            PATH_REPEAT = 5
            for repeat in range(PATH_REPEAT):
                for batch_idx in range(N_BATCHES_STAGE2):
                    batch_start = batch_idx * BATCH_SIZE
                    batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES_STAGE2)
                    for test_data_idx in range(batch_start, batch_end):
                        if test_data_idx >= len(stage2_path_data):
                            break
                        state = stage2_path_data[test_data_idx]
                        prev_state = None
                        prev_triggered = None
                        prev_reward = None
                        for step in range(N_STEPS):
                            legal_actions = []
                            for a in range(agent_isolated.action_dim):
                                ddx, ddy, ddz = agent_isolated.decode_action(a)
                                cand_next = (state[0] + ddx, state[1] + ddy, state[2] + ddz)
                                if is_valid_state(cand_next):
                                    legal_actions.append(a)
                            if not legal_actions:
                                break
                            if random.random() < agent_isolated.epsilon:
                                action = random.choice(legal_actions)
                            else:
                                normalized_state = normalize_state(state)
                                state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
                                with torch.no_grad():
                                    q_values = agent_isolated.model(state_tensor)[0]
                                action = legal_actions[torch.argmax(q_values[legal_actions]).item()]
                            ddx, ddy, ddz = agent_isolated.decode_action(action)
                            next_state = clip_state((state[0] + ddx, state[1] + ddy, state[2] + ddz))
                            triggered = execute_Tr(*next_state)
                            reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                            if target_path.issubset(triggered):
                                reward += 2.0
                            done = (step == N_STEPS - 1)
                            td_error = agent_isolated.store_transition(state, action, reward, next_state, done)
                            priority_stats = agent_isolated.replay_buffer.get_priority_statistics()
                            current_priority = priority_stats['mean_priority'] if priority_stats else 0
                            prioritized_metrics.record_step_metrics(reward, td_error, triggered, target_path, current_priority, None)
                            episode_similarities.append(jaccard_similarity(triggered, target_path))
                            episode_td_errors.append(td_error)
                            if prev_reward is not None:
                                prioritized_metrics.record_action_improvement(reward, prev_reward)
                            prev_state = state
                            prev_triggered = triggered
                            prev_reward = reward
                            state = next_state
                            episode_reward += reward
                    if len(agent_isolated.replay_buffer) >= batch_size:
                        agent_isolated.train(batch_size)
                        global_replay_count += 1
                        if global_replay_count % 2 == 0:
                            agent_isolated.update_target_model()
                if stage1_samples:
                    for batch_idx in range(N_BATCHES_STAGE1):
                        batch_start = batch_idx * BATCH_SIZE
                        batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES_STAGE1)
                        for sample_idx in range(batch_start, batch_end):
                            if sample_idx >= len(stage1_samples):
                                break
                            stage1_state_tuple, _, _, _ = stage1_samples[sample_idx]
                            state = stage1_state_tuple
                            prev_state = None
                            prev_triggered = None
                            prev_reward = None
                            for step in range(N_STEPS):
                                legal_actions = []
                                for a in range(agent_isolated.action_dim):
                                    ddx, ddy, ddz = agent_isolated.decode_action(a)
                                    cand_next = (state[0] + ddx, state[1] + ddy, state[2] + ddz)
                                    if is_valid_state(cand_next):
                                        legal_actions.append(a)
                                if not legal_actions:
                                    break
                                if random.random() < agent_isolated.epsilon:
                                    action = random.choice(legal_actions)
                                else:
                                    normalized_state = normalize_state(state)
                                    state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
                                    with torch.no_grad():
                                        q_values = agent_isolated.model(state_tensor)[0]
                                    action = legal_actions[torch.argmax(q_values[legal_actions]).item()]
                                ddx, ddy, ddz = agent_isolated.decode_action(action)
                                next_state = clip_state((state[0] + ddx, state[1] + ddy, state[2] + ddz))
                                triggered = execute_Tr(*next_state)
                                reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                                reward *= 0.8
                                done = (step == N_STEPS - 1)
                                td_error = agent_isolated.store_transition(state, action, reward, next_state, done)
                                priority_stats = agent_isolated.replay_buffer.get_priority_statistics()
                                current_priority = priority_stats['mean_priority'] if priority_stats else 0
                                prioritized_metrics.record_step_metrics(reward, td_error, triggered, target_path, current_priority, None)
                                episode_similarities.append(jaccard_similarity(triggered, target_path))
                                episode_td_errors.append(td_error)
                                if prev_reward is not None:
                                    prioritized_metrics.record_action_improvement(reward, prev_reward)
                                prev_state = state
                                prev_triggered = triggered
                                prev_reward = reward
                                state = next_state
                                episode_reward += reward
                        if len(agent_isolated.replay_buffer) >= batch_size:
                            agent_isolated.train(batch_size)
                            global_replay_count += 1
                            if global_replay_count % 2 == 0:
                                agent_isolated.update_target_model()
            trained_paths.add(path_idx)
        avg_similarity = np.mean(episode_similarities) if episode_similarities else 0
        avg_td_error = np.mean(episode_td_errors) if episode_td_errors else 0
        priority_stats = agent_isolated.replay_buffer.get_priority_statistics()
        prioritized_metrics.record_episode_metrics(episode, episode_reward, avg_similarity, avg_td_error, agent_isolated.epsilon, "isolated", priority_stats)
        if len(trained_paths) == len(isolated_group):
            break
    return agent_isolated


def append_performance_metrics_to_excel(metrics_collector, filepath, run_number):
    """使用openpyxl直接操作Excel，避免pandas版本问题"""
    avg_td_error = np.mean(metrics_collector.td_errors) if metrics_collector.td_errors else 0
    total_reward = metrics_collector.total_reward
    action_improvement_rate = np.mean(metrics_collector.action_improvements) if metrics_collector.action_improvements else 0
    avg_final_similarity = np.mean(metrics_collector.final_output_similarities) if metrics_collector.final_output_similarities else 0
    avg_episode_reward = np.mean(metrics_collector.episode_rewards) if metrics_collector.episode_rewards else 0
    reward_std = np.std(metrics_collector.episode_rewards) if metrics_collector.episode_rewards else 0
    training_time = metrics_collector.end_time - metrics_collector.start_time if metrics_collector.end_time and metrics_collector.start_time else 0
    avg_memory_usage = np.mean(metrics_collector.episode_memory_usage) if metrics_collector.episode_memory_usage else 0
    per_step_time = training_time / metrics_collector.step_count * 1000 if metrics_collector.step_count > 0 else 0
    avg_priority = np.mean(metrics_collector.priority_statistics) if metrics_collector.priority_statistics else 0
    avg_importance_weight = np.mean(metrics_collector.importance_weights) if metrics_collector.importance_weights else 0

    headers = ['Run', '最终样本平均相似度', '总奖励', '平均回合奖励', '奖励标准差', '平均TD误差',
               '动作改进率', '动作改进率(%)', '总步数', '最终样本数', '训练总时间(秒)',
               '训练总时间(分钟)', '平均内存使用(MB)', '每步时间(ms)', '平均优先级', '平均重要性权重']

    new_row = [
        f"Run {run_number}",
        f"{avg_final_similarity:.4f}",
        f"{total_reward:,.2f}",
        f"{avg_episode_reward:,.4f}",
        f"{reward_std:,.4f}",
        f"{avg_td_error:.4f}",
        f"{action_improvement_rate:.4f}",
        f"{action_improvement_rate * 100:.2f}%",
        f"{metrics_collector.step_count:,}",
        f"{len(metrics_collector.final_output_similarities)}",
        f"{training_time:.2f}",
        f"{training_time / 60:.2f}",
        f"{avg_memory_usage:.2f}",
        f"{per_step_time:.2f}",
        f"{avg_priority:.4f}",
        f"{avg_importance_weight:.4f}"
    ]

    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        if '性能指标' in wb.sheetnames:
            ws = wb['性能指标']
        else:
            ws = wb.create_sheet('性能指标')
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = '性能指标'

    # 如果是新sheet，写入表头
    if ws.max_row == 1:
        # 检查是否已有内容
        has_content = False
        for cell in ws[1]:
            if cell.value is not None:
                has_content = True
                break
        if not has_content:
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)

    # 追加新行
    next_row = ws.max_row + 1
    for col_idx, value in enumerate(new_row, 1):
        ws.cell(row=next_row, column=col_idx, value=value)

    # 设置列宽和格式
    ws.column_dimensions['A'].width = 15
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
        ws.column_dimensions[col].width = 18

    header_font = Font(bold=True, size=11)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(filepath)
    print(f"运行 {run_number} 的性能指标已保存到: {filepath}")


def append_final_samples_to_excel(agent_similar, agent_isolated, similar_group, isolated_group, targetPaths, filepath, run_number):
    """使用openpyxl直接操作Excel，避免pandas版本问题"""
    headers = ['Run', '路径分组', '路径ID', 'dx', 'dy', 'dz', '相似度', '奖励', '触发分支数', '目标分支数', '触发分支', '目标分支']

    new_samples = []
    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_similar.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, reward, sim, triggered in high_reward_samples:
            new_samples.append([
                f"Run {run_number}", '相似路径组', path_idx + 1,
                state_tuple[0], state_tuple[1], state_tuple[2],
                f"{sim:.4f}", f"{reward:.2f}",
                len(triggered), len(target_path),
                str(sorted(triggered)), str(sorted(target_path))
            ])

    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, reward, sim, triggered in high_reward_samples:
            new_samples.append([
                f"Run {run_number}", '孤立路径组', path_idx + 1,
                state_tuple[0], state_tuple[1], state_tuple[2],
                f"{sim:.4f}", f"{reward:.2f}",
                len(triggered), len(target_path),
                str(sorted(triggered)), str(sorted(target_path))
            ])

    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        if '最终样本' in wb.sheetnames:
            ws = wb['最终样本']
        else:
            ws = wb.create_sheet('最终样本')
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = '最终样本'

    # 如果是新sheet，写入表头
    if ws.max_row == 1:
        has_content = False
        for cell in ws[1]:
            if cell.value is not None:
                has_content = True
                break
        if not has_content:
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)

    # 追加新行
    next_row = ws.max_row + 1
    for sample in new_samples:
        for col_idx, value in enumerate(sample, 1):
            ws.cell(row=next_row, column=col_idx, value=value)
        next_row += 1

    # 设置列宽
    column_widths = {'A': 12, 'B': 15, 'C': 12, 'D': 10, 'E': 10, 'F': 10, 'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 40, 'L': 40}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    header_font = Font(bold=True, size=11)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    wb.save(filepath)
    print(f"运行 {run_number} 的最终样本已保存到: {filepath}")


def run_single_experiment(run_number, results_save_dir):
    print(f"\n{'=' * 80}")
    print(f"开始运行第 {run_number} 次实验（优先级DQN）")
    print(f"{'=' * 80}\n")

    prioritized_metrics.reset()
    prioritized_metrics.start_training()

    model_path_similar = os.path.join(results_save_dir, f"run{run_number}_similar.pth")
    model_path_isolated = os.path.join(results_save_dir, f"run{run_number}_isolated.pth")
    path_documents = r"D:\Experiment\CNN\DQNNEW\path_samples_prioritized"

    similar_group, isolated_group = group_paths_by_similarity(targetPaths)
    os.makedirs(path_documents, exist_ok=True)

    samples_exist = True
    for path_idx in similar_group:
        file_path = os.path.join(path_documents, f"prioritized_path{path_idx + 1}.txt")
        if not os.path.exists(file_path):
            samples_exist = False
            break

    if not samples_exist:
        print("生成相似路径样本...")
        generate_samples_for_similar_paths(similar_group, num_total=2000, top_k=200)
    else:
        print("相似路径样本文件已存在，跳过生成...")

    replay_buffer = PrioritizedExperienceReplay(capacity=10000, alpha=0.6, beta_start=0.4, beta_frames=100000)
    state_dim = 3
    action_dim = 30
    agent = PrioritizedDQNAgent(state_dim, action_dim, replay_buffer)

    agent = prioritized_generate_and_train_for_similar_paths(agent, similar_group, path_documents, episodes=500,
                                                             batch_size=32, is_isolated=False)

    os.makedirs(os.path.dirname(model_path_similar), exist_ok=True)
    torch.save({
        'model_state_dict': agent.model.state_dict(),
        'optimizer_state_dict': agent.optimizer.state_dict(),
        'epsilon': agent.epsilon
    }, model_path_similar)

    isolated_samples_exist = True
    for path_idx in isolated_group:
        file_path = os.path.join(path_documents, f"prioritized_path{path_idx + 1}_isolated.txt")
        if not os.path.exists(file_path):
            isolated_samples_exist = False
            break

    if not isolated_samples_exist:
        print("生成孤立路径样本...")
        generate_samples_for_isolated_paths_prioritized(agent, isolated_group, num_total=2000, top_k=200)
    else:
        print("孤立路径样本文件已存在，跳过生成...")

    isolated_replay_buffer = PrioritizedExperienceReplay(capacity=15000, alpha=0.6, beta_start=0.4, beta_frames=100000)
    agent_isolated = PrioritizedDQNAgent(state_dim, action_dim, isolated_replay_buffer)

    try:
        checkpoint = torch.load(model_path_similar, weights_only=True)
        agent_isolated.model.load_state_dict(checkpoint['model_state_dict'])
        agent_isolated.target_model.load_state_dict(checkpoint['model_state_dict'])
        agent_isolated.epsilon = checkpoint.get('epsilon', 0.5)
    except:
        try:
            checkpoint = torch.load(model_path_similar)
            agent_isolated.model.load_state_dict(checkpoint['model_state_dict'])
            agent_isolated.target_model.load_state_dict(checkpoint['model_state_dict'])
            agent_isolated.epsilon = checkpoint.get('epsilon', 0.5)
        except Exception as e:
            print(f"警告: 无法加载相似路径模型权重 - {e}")

    agent_isolated = prioritized_generate_and_train_for_isolated_paths(
        agent_similar=agent, agent_isolated=agent_isolated, similar_group=similar_group,
        isolated_group=isolated_group, path_documents=path_documents, episodes=500,
        batch_size=32, is_isolated=True
    )

    os.makedirs(os.path.dirname(model_path_isolated), exist_ok=True)
    torch.save({
        'model_state_dict': agent_isolated.model.state_dict(),
        'optimizer_state_dict': agent_isolated.optimizer.state_dict(),
        'epsilon': agent_isolated.epsilon
    }, model_path_isolated)

    prioritized_metrics.end_training()

    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, _, sim, triggered in high_reward_samples:
            prioritized_metrics.record_final_output_sample(triggered, target_path)

    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, _, sim, triggered in high_reward_samples:
            prioritized_metrics.record_final_output_sample(triggered, target_path)

    performance_excel_path = os.path.join(results_save_dir, "性能指标.xlsx")
    samples_excel_path = os.path.join(results_save_dir, "最终样本.xlsx")

    append_performance_metrics_to_excel(prioritized_metrics, performance_excel_path, run_number)
    append_final_samples_to_excel(agent, agent_isolated, similar_group, isolated_group, targetPaths, samples_excel_path, run_number)

    avg_similarity = np.mean(prioritized_metrics.final_output_similarities) if prioritized_metrics.final_output_similarities else 0
    training_time = prioritized_metrics.end_time - prioritized_metrics.start_time
    avg_priority = np.mean(prioritized_metrics.priority_statistics) if prioritized_metrics.priority_statistics else 0
    print(f"\n第 {run_number} 次运行完成:")
    print(f"  平均相似度: {avg_similarity:.4f}")
    print(f"  训练时间: {training_time:.2f} 秒")
    print(f"  总步数: {prioritized_metrics.step_count}")
    print(f"  平均优先级: {avg_priority:.4f}")


if __name__ == "__main__":
    results_save_dir = r"D:\Experiment\CNN\DQNNEW\results\prioritized_results"
    os.makedirs(results_save_dir, exist_ok=True)
    NUM_RUNS = 20
    print("=" * 80)
    print(f"开始 {NUM_RUNS} 次优先级DQN实验")
    print("=" * 80)
    for run in range(1, NUM_RUNS + 1):
        try:
            run_single_experiment(run, results_save_dir)
        except Exception as e:
            print(f"\n第 {run} 次运行出错: {str(e)}")
            import traceback
            traceback.print_exc()
            continue
    print("\n" + "=" * 80)
    print(f"所有 {NUM_RUNS} 次运行完成")
    print(f"结果保存在: {results_save_dir}")
    print("=" * 80)