import os
import random
from collections import deque
import numpy as np
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# --- 全局状态范围配置 ---
STATE_MIN_X, STATE_MAX_X = 1, 200
STATE_MIN_Y, STATE_MAX_Y = 1, 200
STATE_MIN_Z, STATE_MAX_Z = 2, 150
def generate_input():
    return [
        random.randint(STATE_MIN_X, STATE_MAX_X),
        random.randint(STATE_MIN_Y, STATE_MAX_Y),
        random.randint(STATE_MIN_Z, STATE_MAX_Z)
    ]



def execute_Tr(dx, dy, dz):
    """
    执行验证规则，返回触发的规则集合
    """
    b = set()

    # 使用全局的质量参数
    global quality_x, quality_y, quality_z
    quality_x = dx
    quality_y = dy
    quality_z = dz

    # 异常类型1：质量参数乘积异常
    if ((quality_y * quality_z) / (quality_x + 1) > 80) != ((quality_y * quality_y * quality_z) / (quality_x + 1) > 80):
        b.add(1)
    if ((quality_y * quality_z) / (quality_x + 1) > 80) != ((quality_y * quality_z * quality_z) / (quality_x + 1) > 80):
        b.add(2)
    if ((quality_y * quality_z) / (quality_x + 1) > 80) != ((quality_y * quality_x * quality_z) / (quality_x + 1) > 80):
        b.add(3)
    if ((quality_y * quality_z) / (quality_x + 1) > 80) != ((quality_y * quality_z) / (quality_x + 1) > 60):
        b.add(4)
    if ((quality_y * quality_z) / (quality_x + 1) > 80) != ((quality_y * quality_z) / (quality_x + 10) > 80):
        b.add(5)
    if ((quality_y * quality_z) / (quality_x + 1) > 80) != ((quality_y * quality_z) / (quality_x + 13) > 80):
        b.add(6)
    if ((quality_y * quality_z) / (quality_x + 1) > 80) != ((quality_y * quality_z * 5) / (quality_x + 1) > 80):
        b.add(7)
    if ((quality_y * quality_z) / (quality_x + 1) > 80) != ((quality_y * quality_z * 2) / (quality_x + 1) > 80):
        b.add(8)
    if ((quality_y * quality_z) / (quality_x + 1) > 80) != ((quality_y * quality_z) / (quality_x + 1) > 40):
        b.add(9)
    if ((quality_y * quality_z) / (quality_x + 1) > 80) != ((quality_y * quality_x) / (quality_x + 1) > 80):
        b.add(10)
    if ((quality_y * quality_z) / (quality_x + 1) > 80) != ((quality_y * quality_y) / (quality_x + 1) > 80):
        b.add(11)
    if ((quality_y * quality_z) / (quality_x + 1) > 80) != ((quality_z * quality_z) / (quality_x + 1) > 80):
        b.add(12)

    # 异常类型2：质量差值异常
    if ((quality_z - quality_x) < 0.4 * quality_y) != ((quality_z - quality_x) < 0.3 * quality_y):
        b.add(13)
    if ((quality_z - quality_x) < 0.4 * quality_y) != ((quality_z - quality_x) < 0.5 * quality_y):
        b.add(14)
    if ((quality_z - quality_x) < 0.4 * quality_y) != ((quality_z - quality_x) < 0.4 * quality_z):
        b.add(15)
    if ((quality_z - quality_x) < 0.4 * quality_y) != ((quality_z - quality_x) < 0.4 * quality_x):
        b.add(16)
    if ((quality_z - quality_x) < 0.4 * quality_y) != ((quality_z * 1.1 - quality_x) < 0.4 * quality_y):
        b.add(17)
    if ((quality_z - quality_x) < 0.4 * quality_y) != ((quality_z * 2 - quality_x) < 0.4 * quality_y):
        b.add(18)
    if ((quality_z - quality_x) < 0.4 * quality_y) != ((quality_z * quality_z - quality_x) < 0.4 * quality_y):
        b.add(19)
    if ((quality_z - quality_x) < 0.4 * quality_y) != ((quality_z * quality_x - quality_x) < 0.4 * quality_y):
        b.add(20)
    if ((quality_z - quality_x) < 0.4 * quality_y) != ((quality_z * quality_y - quality_x) < 0.4 * quality_y):
        b.add(21)
    if ((quality_z - quality_x) < 0.4 * quality_y) != ((quality_z * 1.5 - quality_x) < 0.4 * quality_y):
        b.add(22)

    # 异常类型3：质量立方关系
    if ((quality_x ** 3 + quality_y ** 3) < quality_z ** 2) != ((quality_x ** 2 + quality_y ** 3) < quality_z ** 2):
        b.add(23)
    if ((quality_x ** 3 + quality_y ** 3) < quality_z ** 2) != ((quality_x ** 3 + quality_y ** 2) < quality_z ** 2):
        b.add(24)
    if ((quality_x ** 3 + quality_y ** 3) < quality_z ** 2) != ((quality_x ** 3 + quality_y ** 3) < quality_z ** 1):
        b.add(25)
    if ((quality_x ** 3 + quality_y ** 3) < quality_z ** 2) != ((quality_x ** 3 + quality_y ** 3) < quality_z ** 3):
        b.add(26)
    if ((quality_x ** 3 + quality_y ** 3) < quality_z ** 2) != ((quality_x ** 3 + quality_y ** 3) < quality_z ** 4):
        b.add(27)
    if ((quality_x ** 3 + quality_y ** 3) < quality_z ** 2) != ((quality_x ** 3 + quality_y ** 4) < quality_z ** 2):
        b.add(28)
    if ((quality_x ** 3 + quality_y ** 3) < quality_z ** 2) != ((quality_x ** 3 + quality_x ** 3) < quality_z ** 2):
        b.add(29)
    if ((quality_x ** 3 + quality_y ** 3) < quality_z ** 2) != ((quality_x ** 1 + quality_y ** 3) < quality_z ** 2):
        b.add(30)
    if ((quality_x ** 3 + quality_y ** 3) < quality_z ** 2) != ((quality_x ** 3 + quality_y ** 1) < quality_z ** 2):
        b.add(31)
    if ((quality_x ** 3 + quality_y ** 3) < quality_z ** 2) != (
            ((quality_x ** 3) * 2 + quality_y ** 3) < quality_z ** 2):
        b.add(32)
    if ((quality_x ** 3 + quality_y ** 3) < quality_z ** 2) != (
            (quality_x ** 3 + (quality_y ** 3) * 2) < quality_z ** 2):
        b.add(33)
    if ((quality_x ** 3 + quality_y ** 3) < quality_z ** 2) != ((quality_y ** 3 + quality_y ** 3) < quality_z ** 2):
        b.add(34)
    if ((quality_x ** 3 + quality_y ** 3) < quality_z ** 2) != (
            (quality_x ** 3 + quality_y ** 3) < (quality_z ** 2) * 2):
        b.add(35)
    if ((quality_x ** 3 + quality_y ** 3) < quality_z ** 2) != (
            (quality_x ** 3 + quality_y ** 3) < (quality_x ** 2) * 2):
        b.add(36)
    if ((quality_x ** 3 + quality_y ** 3) < quality_z ** 2) != (
            (quality_x ** 3 + quality_y ** 3) < (quality_y ** 2) * 2):
        b.add(37)

    # 异常类型6：质量同步性检查
    if (abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1) != (
            abs(quality_x % 2 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1):
        b.add(38)
    if (abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1) != (
            abs(quality_x % 3 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1):
        b.add(39)
    if (abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1) != (
            abs(quality_x % 1 - quality_y % 2) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1):
        b.add(40)
    if (abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1) != (
            abs(quality_x % 1 - quality_y % 3) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1):
        b.add(41)
    if (abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1) != (
            abs(quality_x % 1 - quality_y % 5) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1):
        b.add(42)
    if (abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1) != (
            abs(quality_x % 5 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1):
        b.add(43)
    if (abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1) != (
            abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 2 - quality_z % 1) < 0.1):
        b.add(44)
    if (abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1) != (
            abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 3 - quality_z % 1) < 0.1):
        b.add(45)
    if (abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1) != (
            abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 5 - quality_z % 1) < 0.1):
        b.add(46)
    if (abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1) != (
            abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 2) < 0.1):
        b.add(47)
    if (abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1) != (
            abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 3) < 0.1):
        b.add(48)
    if (abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1) != (
            abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 5) < 0.1):
        b.add(49)
    if (abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1) != (
            abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 4 - quality_z % 1) < 0.1):
        b.add(50)
    if (abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1) != (
            abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 4) < 0.1):
        b.add(51)
    if (abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1) != (
            abs(quality_x % 4 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1):
        b.add(52)
    if (abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1) != (
            abs(quality_x % 1 - quality_y % 6) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1):
        b.add(53)
    if (abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1) != (
            abs(quality_x % 6 - quality_y % 1) < 0.1 and abs((quality_y * 2) % 1 - quality_z % 1) < 0.1):
        b.add(54)
    if (abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 1 - quality_z % 1) < 0.1) != (
            abs(quality_x % 1 - quality_y % 1) < 0.1 and abs(quality_y % 6 - (quality_z * 2) % 1) < 0.1):
        b.add(55)

    # 其他复杂检查逻辑
    if (quality_x * quality_y * quality_z > 500000 and (quality_x + quality_y + quality_z) / 3 < 85) != (
            quality_x * quality_y * quality_z > 500000 and (quality_x + quality_y + quality_z) / 2 < 85):
        b.add(56)
    if (quality_x * quality_y * quality_z > 500000 and (quality_x + quality_y + quality_z) / 3 < 85) != (
            quality_x * quality_y * quality_z > 500000 and (quality_x + quality_y + quality_z) / 4 < 85):
        b.add(57)
    if (quality_x * quality_y * quality_z > 500000 and (quality_x + quality_y + quality_z) / 3 < 85) != (
            quality_x * quality_y * quality_z > 500000 and (quality_x * 2 + quality_y + quality_z) / 3 < 85):
        b.add(58)
    if (quality_x * quality_y * quality_z > 500000 and (quality_x + quality_y + quality_z) / 3 < 85) != (
            quality_x * quality_y * quality_z > 500000 and (quality_x + quality_y * 2 + quality_z) / 3 < 85):
        b.add(59)
    if (quality_x * quality_y * quality_z > 500000 and (quality_x + quality_y + quality_z) / 3 < 85) != (
            quality_x * quality_y * quality_z > 500000 and (quality_x + quality_y - quality_z) / 3 < 85):
        b.add(60)
    if (quality_x * quality_y * quality_z > 500000 and (quality_x + quality_y + quality_z) / 3 < 85) != (
            quality_x * quality_y * quality_z > 500000 and (quality_x + quality_y * quality_z * 2) / 3 < 85):
        b.add(61)
    if (quality_x * quality_y * quality_z > 500000 and (quality_x + quality_y + quality_z) / 3 < 85) != (
            quality_x * quality_y * quality_z > 500000 and (quality_x + quality_y + quality_x) / 3 < 85):
        b.add(62)
    if (quality_x * quality_y * quality_z > 500000 and (quality_x + quality_y + quality_z) / 3 < 85) != (
            quality_x * quality_y * quality_z > 500000 and (quality_x + quality_y + quality_y) / 3 < 85):
        b.add(63)
    if (quality_x * quality_y * quality_z > 500000 and (quality_x + quality_y + quality_z) / 3 < 85) != (
            quality_x * quality_y * quality_z > 600000 and (quality_x + quality_y + quality_z) / 3 < 85):
        b.add(64)
    if (quality_x * quality_y * quality_z > 500000 and (quality_x + quality_y + quality_z) / 3 < 85) != (
            quality_x * quality_y * quality_z * 2 > 500000 and (quality_x + quality_y + quality_z) / 3 < 85):
        b.add(65)
    if (quality_x * quality_y * quality_z > 500000 and (quality_x + quality_y + quality_z) / 3 < 85) != (
            quality_y * quality_y * quality_z > 500000 and (quality_x + quality_y + quality_z) / 3 < 85):
        b.add(66)
    if (quality_x * quality_y * quality_z > 500000 and (quality_x + quality_y + quality_z) / 3 < 85) != (
            quality_z * quality_y * quality_z > 500000 and (quality_x + quality_y + quality_z) / 3 < 85):
        b.add(67)
    if (quality_x * quality_y * quality_z > 500000 and (quality_x + quality_y + quality_z) / 3 < 85) != (
            quality_x * quality_x * quality_z > 500000 and (quality_x + quality_y + quality_z) / 3 < 85):
        b.add(68)
    if (quality_x * quality_y * quality_z > 500000 and (quality_x + quality_y + quality_z) / 3 < 85) != (
            quality_x * quality_y * quality_x > 500000 and (quality_x + quality_y + quality_z) / 3 < 85):
        b.add(69)
    if (quality_x * quality_y * quality_z > 500000 and (quality_x + quality_y + quality_z) / 3 < 85) != (
            quality_x * quality_z * quality_z > 500000 and (quality_x + quality_y + quality_z) / 3 < 85):
        b.add(70)
    if (quality_x * quality_y * quality_z > 500000 and (quality_x + quality_y + quality_z) / 3 < 85) != (
            quality_x * quality_y * quality_y > 500000 and (quality_x + quality_y + quality_z) / 3 < 85):
        b.add(71)

    return b


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0


targetPaths = [
    {1, 2, 3, 4, 7, 8, 9, 10, 11, 19, 20, 21, 27, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 49, 50, 51, 52, 53, 54, 55,
     56, 58, 59, 61, 62, 63, 64, 67, 68, 70},
    {1, 2, 3, 4, 7, 8, 9, 12, 18, 19, 20, 21, 22, 27, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53,
     54, 55, 56, 58, 59, 61, 64, 66, 71},
    {1, 2, 3, 4, 7, 8, 9, 12, 13, 17, 18, 19, 20, 21, 22, 26, 27, 40, 41, 42, 44, 45, 46, 48, 49, 50, 51, 52, 53, 55,
     56, 58, 59, 61, 64, 66, 69, 71},
    {5, 6, 10, 11, 13, 17, 18, 19, 20, 21, 22, 26, 27, 38, 39, 40, 42, 43, 44, 46, 49, 50, 52, 53, 54, 55, 56, 58, 59,
     61, 64, 66, 69, 71},
    {1, 2, 3, 7, 12, 17, 18, 19, 20, 21, 22, 26, 27, 30, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 49, 50, 51, 52, 53, 54,
     55, 65, 68, 70},
    {1, 2, 3, 7, 8, 9, 12, 17, 18, 19, 20, 21, 22, 26, 27, 38, 39, 40, 41, 42, 44, 45, 46, 47, 48, 50, 51, 52, 53, 54,
     55, 57, 60, 63},
    {1, 2, 3, 4, 7, 8, 9, 12, 23, 24, 26, 27, 30, 31, 34, 35, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51,
     52, 53, 54, 55},
    {16, 18, 19, 20, 21, 22, 27, 38, 40, 41, 42, 43, 44, 45, 46, 48, 49, 50, 51, 52, 53, 54, 55, 56, 58, 59, 61, 63, 64,
     68, 69, 70},
    {1, 2, 3, 4, 7, 8, 9, 12, 14, 15, 16, 26, 27, 38, 40, 41, 42, 43, 44, 45, 46, 48, 49, 50, 51, 52, 53, 54, 55, 57,
     60, 62, 63},
    {5, 6, 10, 13, 15, 16, 18, 19, 20, 21, 22, 27, 29, 31, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53,
     54, 55},
    {1, 2, 3, 7, 8, 9, 12, 25, 28, 29, 32, 33, 36, 37, 38, 40, 41, 42, 43, 44, 45, 46, 47, 48, 50, 51, 52, 53, 54, 55,
     70}
]


class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 2000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(targetPaths)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }


def compute_robustness(state, path):
    base = execute_Tr(*state)
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                        np.clip(state[0] + dx, STATE_MIN_X, STATE_MAX_X),
                        np.clip(state[1] + dy, STATE_MIN_Y, STATE_MAX_Y),
                        np.clip(state[2] + dz, STATE_MIN_Z, STATE_MAX_Z)
                    ])
                n_trig = execute_Tr(*neighbor)
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = targetPaths[target_path_idx]
    samples = []
    attempts = 0

    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(STATE_MIN_X, STATE_MAX_X),
            random.randint(STATE_MIN_Y, STATE_MAX_Y),
            random.randint(STATE_MIN_Z, STATE_MAX_Z)
        ])
        triggered = execute_Tr(*state)

        if not triggered:
            continue

        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)

        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }

        samples.append(sample_data)

    return samples


def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample[
                'robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)

        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]

    return selected_samples


def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = targetPaths[target_path_idx]
        samples = []
        attempts = 0

        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(STATE_MIN_X, STATE_MAX_X),
                random.randint(STATE_MIN_Y, STATE_MAX_Y),
                random.randint(STATE_MIN_Z, STATE_MAX_Z)
            ])
            triggered = execute_Tr(*state)

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)

            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)

        random.shuffle(samples)
        return samples[:config.top_k_samples]

    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")

        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)


def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}

    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates

    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}

        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config,
                                                         shared_candidates[path_idx])

            strategy_results[path_idx] = samples

        results[strategy_name] = strategy_results

    return results


def analyze_fitness_values(results, config):
    analysis_results = {}

    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []

        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])

        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }

        all_scores = []

        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])

        scores_array = np.array(all_scores)

        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)

        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })

        analysis_results[strategy_name] = analysis

    return analysis_results


def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []

    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)

    df = pd.DataFrame(df_data)
    return df, analysis_results


def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []

    print(f"Starting {num_runs} experiments...")

    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")

        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)

        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)

        print(f"Experiment {run_idx + 1} completed.")

    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df


def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')

    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width

    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []

    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]

        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)

    stats_df = pd.DataFrame(stats_data)

    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Results saved to: {output_path}")


def main():
    results_df = run_multiple_experiments(num_runs=20)

    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")

    save_results_to_excel(results_df, output_path)

    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)

    return results_df, output_path


if __name__ == "__main__":
    results, output_file = main()