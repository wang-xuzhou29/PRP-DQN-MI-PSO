import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
import numpy as np
import random
import time
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# === 实验配置 ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'MIN_X': 1, 'MAX_X': 200,
    'MIN_Y': 1, 'MAX_Y': 200,
    'MIN_Z': 2, 'MAX_Z': 150,
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 5,
    'NUM_ROUNDS': 5,
    'REPLAY_BATCH_SIZE': 64,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'LEARNING_RATE': 3e-4,
    'NUM_RUNS': 20,  # 20次运行
    'TOP_K_SAMPLES': 20,
    'REPLAY_BUFFER_CAPACITY': 20000,  # 每个路径的回放缓冲区容量
    'TARGET_PATHS': [
    {1, 2, 3, 4, 7, 8, 9, 10, 11, 19, 20, 21, 27, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 49, 50, 51, 52, 53, 54, 55,
     56, 58, 59, 61, 62, 63, 64, 67, 68, 70},
    {1, 2, 3, 4, 7, 8, 9, 12, 18, 19, 20, 21, 22, 27, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53,
     54, 55, 56, 58, 59, 61, 64, 66, 71},
    {1, 2, 3, 4, 7, 8, 9, 12, 13, 17, 18, 19, 20, 21, 22, 26, 27, 40, 41, 42, 44, 45, 46, 48, 49, 50, 51, 52, 53, 55,
     56, 58, 59, 61, 64, 66, 69, 71},
    {5, 6, 10, 11, 13, 17, 18, 19, 20, 21, 22, 26, 27, 38, 39, 40, 42, 43, 44, 46, 49, 50, 52, 53, 54, 55, 56, 58, 59,
     61, 64, 66, 69, 71},
    {1, 2, 3, 7, 12, 17, 18, 19, 20, 21, 22, 26, 27, 30, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 49, 50, 51, 52, 53, 54,
     55, 65, 68, 70},
    {1, 2, 3, 7, 8, 9, 12, 17, 18, 19, 20, 21, 22, 26, 27, 38, 39, 40, 41, 42, 44, 45, 46, 47, 48, 50, 51, 52, 53, 54,
     55, 57, 60, 63},
    {1, 2, 3, 4, 7, 8, 9, 12, 23, 24, 26, 27, 30, 31, 34, 35, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51,
     52, 53, 54, 55},
    {16, 18, 19, 20, 21, 22, 27, 38, 40, 41, 42, 43, 44, 45, 46, 48, 49, 50, 51, 52, 53, 54, 55, 56, 58, 59, 61, 63, 64,
     68, 69, 70},
    {1, 2, 3, 4, 7, 8, 9, 12, 14, 15, 16, 26, 27, 38, 40, 41, 42, 43, 44, 45, 46, 48, 49, 50, 51, 52, 53, 54, 55, 57,
     60, 62, 63},
    {5, 6, 10, 13, 15, 16, 18, 19, 20, 21, 22, 27, 29, 31, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53,
     54, 55},
    {1, 2, 3, 7, 8, 9, 12, 25, 28, 29, 32, 33, 36, 37, 38, 40, 41, 42, 43, 44, 45, 46, 47, 48, 50, 51, 52, 53, 54, 55,
     70}
    ],
}


# === 辅助函数 ===
def get_bounds():
    mins = np.array([EXPERIMENT_CONFIG['MIN_X'], EXPERIMENT_CONFIG['MIN_Y'], EXPERIMENT_CONFIG['MIN_Z']])
    maxs = np.array([EXPERIMENT_CONFIG['MAX_X'], EXPERIMENT_CONFIG['MAX_Y'], EXPERIMENT_CONFIG['MAX_Z']])
    return mins, maxs

def clip_state(state):
    mins, maxs = get_bounds()
    return np.clip(state, mins, maxs)

def denormalize_state(normalized_state):
    """将归一化状态还原为原始状态"""
    mins, maxs = get_bounds()
    return normalized_state * (maxs - mins) / 2 + (mins + maxs) / 2


def coverage_similarity(triggered, target_path):
    """
    相似度计算: 交集 / 目标路径大小
    """
    if triggered is None:
        return 0.0
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0

    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)


def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']

    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']

    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']

    return reward


# === 测试需求执行函数 ===
def execute_Tr(x, y, z):
    """执行测试需求，返回触发的分支集合"""
    # 初始化分支覆盖数组
    b = [0] * 71  # 根据分支数量调整大小

    # --- 分支 1-14 ---
    b = set()

    # 异常类型1：质量参数乘积异常
    if ((y * z) / (x + 1) > 80) != ((y * y * z) / (x + 1) > 80):
        b.add(1)
    if ((y * z) / (x + 1) > 80) != ((y * z * z) / (x + 1) > 80):
        b.add(2)
    if ((y * z) / (x + 1) > 80) != ((y * x * z) / (x + 1) > 80):
        b.add(3)
    if ((y * z) / (x + 1) > 80) != ((y * z) / (x + 1) > 60):
        b.add(4)
    if ((y * z) / (x + 1) > 80) != ((y * z) / (x + 10) > 80):
        b.add(5)
    if ((y * z) / (x + 1) > 80) != ((y * z) / (x + 13) > 80):
        b.add(6)
    if ((y * z) / (x + 1) > 80) != ((y * z * 5) / (x + 1) > 80):
        b.add(7)
    if ((y * z) / (x + 1) > 80) != ((y * z * 2) / (x + 1) > 80):
        b.add(8)
    if ((y * z) / (x + 1) > 80) != ((y * z) / (x + 1) > 40):
        b.add(9)
    if ((y * z) / (x + 1) > 80) != ((y * x) / (x + 1) > 80):
        b.add(10)
    if ((y * z) / (x + 1) > 80) != ((y * y) / (x + 1) > 80):
        b.add(11)
    if ((y * z) / (x + 1) > 80) != ((z * z) / (x + 1) > 80):
        b.add(12)

    # 异常类型2：质量差值异常
    if ((z - x) < 0.4 * y) != ((z - x) < 0.3 * y):
        b.add(13)
    if ((z - x) < 0.4 * y) != ((z - x) < 0.5 * y):
        b.add(14)
    if ((z - x) < 0.4 * y) != ((z - x) < 0.4 * z):
        b.add(15)
    if ((z - x) < 0.4 * y) != ((z - x) < 0.4 * x):
        b.add(16)
    if ((z - x) < 0.4 * y) != ((z * 1.1 - x) < 0.4 * y):
        b.add(17)
    if ((z - x) < 0.4 * y) != ((z * 2 - x) < 0.4 * y):
        b.add(18)
    if ((z - x) < 0.4 * y) != ((z * z - x) < 0.4 * y):
        b.add(19)
    if ((z - x) < 0.4 * y) != ((z * x - x) < 0.4 * y):
        b.add(20)
    if ((z - x) < 0.4 * y) != ((z * y - x) < 0.4 * y):
        b.add(21)
    if ((z - x) < 0.4 * y) != ((z * 1.5 - x) < 0.4 * y):
        b.add(22)

    # 异常类型3：质量立方关系
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2 + y ** 3) < z ** 2):
        b.add(23)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 2) < z ** 2):
        b.add(24)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 1):
        b.add(25)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 3):
        b.add(26)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 4):
        b.add(27)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 4) < z ** 2):
        b.add(28)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + x ** 3) < z ** 2):
        b.add(29)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 1 + y ** 3) < z ** 2):
        b.add(30)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 1) < z ** 2):
        b.add(31)
    if ((x ** 3 + y ** 3) < z ** 2) != (
            ((x ** 3) * 2 + y ** 3) < z ** 2):
        b.add(32)
    if ((x ** 3 + y ** 3) < z ** 2) != (
            (x ** 3 + (y ** 3) * 2) < z ** 2):
        b.add(33)
    if ((x ** 3 + y ** 3) < z ** 2) != ((y ** 3 + y ** 3) < z ** 2):
        b.add(34)
    if ((x ** 3 + y ** 3) < z ** 2) != (
            (x ** 3 + y ** 3) < (z ** 2) * 2):
        b.add(35)
    if ((x ** 3 + y ** 3) < z ** 2) != (
            (x ** 3 + y ** 3) < (x ** 2) * 2):
        b.add(36)
    if ((x ** 3 + y ** 3) < z ** 2) != (
            (x ** 3 + y ** 3) < (y ** 2) * 2):
        b.add(37)

    # 异常类型6：质量同步性检查
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 2 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1):
        b.add(38)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 3 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1):
        b.add(39)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 2) < 0.1 and abs(y % 1 - z % 1) < 0.1):
        b.add(40)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 3) < 0.1 and abs(y % 1 - z % 1) < 0.1):
        b.add(41)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 5) < 0.1 and abs(y % 1 - z % 1) < 0.1):
        b.add(42)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 5 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1):
        b.add(43)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 1) < 0.1 and abs(y % 2 - z % 1) < 0.1):
        b.add(44)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 1) < 0.1 and abs(y % 3 - z % 1) < 0.1):
        b.add(45)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 1) < 0.1 and abs(y % 5 - z % 1) < 0.1):
        b.add(46)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 2) < 0.1):
        b.add(47)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 3) < 0.1):
        b.add(48)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 5) < 0.1):
        b.add(49)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 1) < 0.1 and abs(y % 4 - z % 1) < 0.1):
        b.add(50)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 4) < 0.1):
        b.add(51)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 4 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1):
        b.add(52)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 6) < 0.1 and abs(y % 1 - z % 1) < 0.1):
        b.add(53)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 6 - y % 1) < 0.1 and abs((y * 2) % 1 - z % 1) < 0.1):
        b.add(54)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 1) < 0.1 and abs(y % 6 - (z * 2) % 1) < 0.1):
        b.add(55)

    # 其他复杂检查逻辑
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * y * z > 500000 and (x + y + z) / 2 < 85):
        b.add(56)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * y * z > 500000 and (x + y + z) / 4 < 85):
        b.add(57)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * y * z > 500000 and (x * 2 + y + z) / 3 < 85):
        b.add(58)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * y * z > 500000 and (x + y * 2 + z) / 3 < 85):
        b.add(59)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * y * z > 500000 and (x + y - z) / 3 < 85):
        b.add(60)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * y * z > 500000 and (x + y * z * 2) / 3 < 85):
        b.add(61)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * y * z > 500000 and (x + y + x) / 3 < 85):
        b.add(62)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * y * z > 500000 and (x + y + y) / 3 < 85):
        b.add(63)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * y * z > 600000 and (x + y + z) / 3 < 85):
        b.add(64)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * y * z * 2 > 500000 and (x + y + z) / 3 < 85):
        b.add(65)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            y * y * z > 500000 and (x + y + z) / 3 < 85):
        b.add(66)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            z * y * z > 500000 and (x + y + z) / 3 < 85):
        b.add(67)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * x * z > 500000 and (x + y + z) / 3 < 85):
        b.add(68)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * y * x > 500000 and (x + y + z) / 3 < 85):
        b.add(69)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * z * z > 500000 and (x + y + z) / 3 < 85):
        b.add(70)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * y * y > 500000 and (x + y + z) / 3 < 85):
        b.add(71)
    # 返回触发的分支集合
    triggered = set()
    for i, val in enumerate(b):
        if val > 0:
            triggered.add(i + 1)  # 分支编号从1开始
    return triggered


# === DQN网络 ===
class DQNNetwork(nn.Module):
    def __init__(self, action_size=30):
        super(DQNNetwork, self).__init__()
        hidden_dim = EXPERIMENT_CONFIG['HIDDEN_DIM']

        self.fc1 = nn.Linear(3, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)
        self.output = nn.Linear(hidden_dim, action_size)

    def forward(self, x):
        x = torch.relu(self.fc1(x))
        x = torch.relu(self.fc2(x))
        x = torch.relu(self.fc3(x))
        return self.output(x)


# === 每个路径的回放缓冲区 ===
class PathReplayBuffer:
    """每个路径独立的回放缓冲区"""

    def __init__(self, path_idx, capacity=20000):
        self.path_idx = path_idx
        self.capacity = capacity
        self.buffer = deque(maxlen=capacity)
        self.similarities = deque(maxlen=capacity)  # 存储相似度

    def push(self, state, action, reward, next_state, done, similarity):
        self.buffer.append((state, action, reward, next_state, done))
        self.similarities.append(similarity)

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return None

        batch = random.sample(self.buffer, batch_size)
        state, action, reward, next_state, done = map(np.stack, zip(*batch))

        return (
            torch.FloatTensor(state).to(device),
            torch.LongTensor(action).to(device),
            torch.FloatTensor(reward).to(device),
            torch.FloatTensor(next_state).to(device),
            torch.BoolTensor(done).to(device)
        )

    def get_top_k(self, k=20):
        """获取每个路径的Top-K样本"""
        if len(self.buffer) == 0:
            return []

        # 将buffer和similarities配对
        samples_with_sim = list(zip(self.buffer, self.similarities))
        samples_with_sim.sort(key=lambda x: x[1], reverse=True)

        # 取Top-K
        top_k = samples_with_sim[:k]

        results = []
        for (state, _, _, _, _), similarity in top_k:
            original_state = denormalize_state(state)
            original_state_int = np.round(original_state).astype(int)

            results.append({
                'state': original_state_int,
                'similarity': similarity,
                'triggered': execute_Tr(*original_state_int)
            })

        return results

    def __len__(self):
        return len(self.buffer)


# === 改进的DQN智能体(每个路径独立缓冲区) ===
class ImprovedDQNAgent:
    def __init__(self, num_paths, action_size=30):
        self.action_size = action_size
        self.num_paths = num_paths
        self.epsilon = 0.9
        self.epsilon_min = 0.1
        self.epsilon_decay = 0.995

        self.q_network = DQNNetwork(action_size).to(device)
        self.target_network = DQNNetwork(action_size).to(device)

        lr = EXPERIMENT_CONFIG['LEARNING_RATE']
        self.optimizer = optim.Adam(self.q_network.parameters(), lr=lr)

        # 为每个路径创建独立的回放缓冲区
        capacity = EXPERIMENT_CONFIG['REPLAY_BUFFER_CAPACITY']
        self.replay_buffers = {}
        for path_idx in range(num_paths):
            self.replay_buffers[path_idx] = PathReplayBuffer(path_idx, capacity)

        self.replay_train_count = 0
        self.update_target_network()

    def discrete_to_action_delta(self, action_idx):
        # 定义每个维度的离散变化值
        delta_values = [5, 3, 2, 1, 0.5, -0.5, -1, -2, -3, -5]

        if action_idx >= 30:
            action_idx = action_idx % 30

        dim = action_idx // 10
        delta_idx = action_idx % 10
        delta = delta_values[delta_idx]

        action_delta = np.zeros(3)
        action_delta[dim] = delta

        return action_delta

    def get_action(self, state):
        mins, maxs = get_bounds()
        normalized_state = (state - (mins + maxs) / 2) / ((maxs - mins) / 2)

        if random.random() < self.epsilon:
            action_idx = random.randint(0, self.action_size - 1)
        else:
            state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)
            with torch.no_grad():
                q_values = self.q_network(state_tensor)
                action_idx = q_values.argmax().item()

        action_delta = self.discrete_to_action_delta(action_idx)
        return action_delta, action_idx

    def store_experience(self, path_idx, state, action_idx, reward, next_state, done, similarity):
        """存储经验到对应路径的缓冲区"""
        mins, maxs = get_bounds()
        normalized_state = (state - (mins + maxs) / 2) / ((maxs - mins) / 2)
        normalized_next_state = (next_state - (mins + maxs) / 2) / ((maxs - mins) / 2)

        self.replay_buffers[path_idx].push(
            normalized_state, action_idx, reward,
            normalized_next_state, done, similarity
        )

    def replay_train(self, path_idx):
        """从指定路径的缓冲区训练"""
        batch_size = EXPERIMENT_CONFIG['REPLAY_BATCH_SIZE']
        batch = self.replay_buffers[path_idx].sample(batch_size)

        if batch is None:
            return

        states, actions, rewards, next_states, dones = batch

        current_q_values = self.q_network(states).gather(1, actions.unsqueeze(1))

        with torch.no_grad():
            next_q_values = self.target_network(next_states).max(1)[0]
            target_q_values = rewards + (0.99 * next_q_values * ~dones)

        loss = F.mse_loss(current_q_values.squeeze(), target_q_values)

        self.optimizer.zero_grad()
        loss.backward()
        torch.nn.utils.clip_grad_norm_(self.q_network.parameters(), 1.0)
        self.optimizer.step()

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

        self.replay_train_count += 1

        if self.replay_train_count % 2 == 0:
            self.update_target_network()
            print(f"    -> 更新目标网络 (第 {self.replay_train_count} 次训练)")

    def update_target_network(self):
        self.target_network.load_state_dict(self.q_network.state_dict())

    def get_all_top_k(self, k=20):
        """获取所有路径的Top-K样本"""
        results = {}
        for path_idx in range(self.num_paths):
            results[path_idx] = self.replay_buffers[path_idx].get_top_k(k)
        return results

    def get_buffer_stats(self):
        """获取缓冲区统计信息"""
        stats = {}
        for path_idx in range(self.num_paths):
            stats[path_idx] = len(self.replay_buffers[path_idx])
        return stats


# === 性能指标计算 ===
def calculate_run_performance(run_idx, dqn_results, training_time, total_steps, update_count, agent):
    """计算单次运行的性能指标"""
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    # 1. 总奖励 (Total Reward)
    total_reward = 0
    # 2. 平均奖励 (Average Reward)
    average_reward = 0
    # 5. 收敛性 (Convergence)
    convergence = 0
    # 12. 环境适应性 (Environment Adaptability)
    environment_adaptability = 0
    # 13. 泛化能力 (Generalization Ability)
    generalization_ability = 0
    # 15. 计算效率 (Computational Efficiency)
    computational_efficiency = 0
    # 16. 策略更新频率 (Policy Update Frequency)
    policy_update_frequency = 0

    # 收集所有相似度
    all_similarities = []

    # 计算各项指标
    total_samples = 0
    all_rewards = []

    for path_idx in range(num_paths):
        samples = dqn_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']

            total_reward += reward
            all_rewards.append(reward)
            all_similarities.append(similarity)
            total_samples += 1

    # 1. 总奖励
    total_reward = total_reward

    # 2. 平均奖励
    if total_samples > 0:
        average_reward = total_reward / total_samples

    # 5. 收敛性（平均相似度）
    if all_similarities:
        convergence = np.mean(all_similarities)

    # 12. 环境适应性（相似度的稳定性）
    if len(all_similarities) > 1:
        environment_adaptability = 1 / (np.std(all_similarities) + 1e-8)

    # 13. 泛化能力（平均相似度）
    generalization_ability = convergence

    # 15. 计算效率（步数/秒）
    if training_time > 0:
        computational_efficiency = total_steps / training_time

    # 16. 策略更新频率
    if training_time > 0:
        policy_update_frequency = update_count / training_time

    # 相似度统计
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0

    return {
        'Run': run_idx + 1,

        # 性能指标
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),

        # 相似度统计
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }


# === Excel导出功能 ===
def export_to_excel(all_dqn_results, all_performance_data, target_paths, output_path="DQN_20次运行结果.xlsx"):
    """将20次运行的DQN结果导出到Excel"""
    print("\n正在导出Excel...")

    # 准备数据
    all_dqn_summary_data = []
    all_dqn_detailed_data = []

    # 遍历每次运行
    for run_idx, (dqn_results, performance_data) in enumerate(zip(all_dqn_results, all_performance_data)):
        # ===== Sheet1: DQN路径摘要 =====
        dqn_summary_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = dqn_results[path_idx]

            if len(samples) == 0:
                dqn_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue

            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = 'Yes' if perfect_count > 0 else 'No'

            dqn_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })
        all_dqn_summary_data.extend(dqn_summary_data)

        # ===== Sheet2: DQN详细样本数据 =====
        dqn_detailed_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = dqn_results[path_idx]

            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']

                dqn_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Sample ID': sample_idx + 1,
                    'X': int(state[0]),
                    'Y': int(state[1]),
                    'Z': int(state[2]),
                    'Similarity': round(similarity, 4),
                    'Perfect Coverage': 'Yes' if abs(similarity - 1.0) < 0.001 else 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path))),
                    'Triggered Rules': ', '.join(map(str, sorted(triggered))),
                    'Hit Rule Count': len(target_path.intersection(triggered)),
                    'Target Rule Count': len(target_path)
                })
        all_dqn_detailed_data.extend(dqn_detailed_data)

    # 创建Excel文件
    dqn_summary_df = pd.DataFrame(all_dqn_summary_data)
    dqn_detailed_df = pd.DataFrame(all_dqn_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet1: DQN路径摘要
        dqn_summary_df.to_excel(writer, sheet_name='DQN路径摘要', index=False)

        # Sheet2: DQN详细样本数据
        dqn_detailed_df.to_excel(writer, sheet_name='DQN详细样本数据', index=False)

        # Sheet3: 性能指标
        selected_columns = [
            'Run',
            'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df_selected = performance_df[selected_columns]
        performance_df_selected.to_excel(writer, sheet_name='性能指标', index=False)

        # 格式化Excel
        workbook = writer.book

        # 定义样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        perfect_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # 绿色

        # === Sheet1 格式化 ===
        ws1 = writer.sheets['DQN路径摘要']
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 高亮完美覆盖的行
        for row_idx in range(2, ws1.max_row + 1):
            if ws1.cell(row_idx, 9).value == 'Yes':  # 第9列是Perfect Coverage
                for col_idx in range(1, ws1.max_column + 1):
                    ws1.cell(row_idx, col_idx).fill = perfect_fill

        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 15
        ws1.column_dimensions['F'].width = 15
        ws1.column_dimensions['G'].width = 15
        ws1.column_dimensions['H'].width = 15
        ws1.column_dimensions['I'].width = 15
        ws1.column_dimensions['J'].width = 50

        # === Sheet2 格式化 ===
        ws2 = writer.sheets['DQN详细样本数据']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 15
        ws2.column_dimensions['I'].width = 40
        ws2.column_dimensions['J'].width = 40
        ws2.column_dimensions['K'].width = 15
        ws2.column_dimensions['L'].width = 15

        # === Sheet3 格式化 ===
        ws3 = writer.sheets['性能指标']
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 设置列宽
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        for col in columns:
            ws3.column_dimensions[col].width = 18

    print(f"Excel已保存: {output_path}")
    print(f"  - Sheet1: DQN路径摘要 ({len(all_dqn_summary_data)} 行)")
    print(f"  - Sheet2: DQN详细样本数据 ({len(all_dqn_detailed_data)} 行)")
    print(f"  - Sheet3: 性能指标 ({len(all_performance_data)} 行)")


# === DQN训练工作流(每个路径独立缓冲区) ===
def train_dqn_workflow():
    print("=" * 80)
    print("DQN训练 (每个路径独立缓冲区)")
    print("相似度: 交集 / 目标路径大小")
    print(
        f"训练配置: 每个路径 {EXPERIMENT_CONFIG['NUM_ROUNDS']} 轮,每轮 {EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']} 步")
    print(f"回放缓冲区: 每个路径独立,容量={EXPERIMENT_CONFIG['REPLAY_BUFFER_CAPACITY']}")
    print("=" * 80)

    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    # 初始化智能体(路径数量)
    agent = ImprovedDQNAgent(num_paths=num_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n生成初始样本: 每个路径 {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']} 个")
    path_samples = {}
    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            state = np.array([
                random.randint(EXPERIMENT_CONFIG['MIN_X'], EXPERIMENT_CONFIG['MAX_X']),
                random.randint(EXPERIMENT_CONFIG['MIN_Y'], EXPERIMENT_CONFIG['MAX_Y']),
                random.randint(EXPERIMENT_CONFIG['MIN_Z'], EXPERIMENT_CONFIG['MAX_Z'])
            ], dtype=np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  路径 {path_idx + 1}/{num_paths}: 生成 {len(samples)} 个样本")

    # 训练参数
    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size
    num_rounds = EXPERIMENT_CONFIG['NUM_ROUNDS']

    print(f"\n训练参数:")
    print(f"  - 批次大小: {batch_size}")
    print(f"  - 每个路径的批次数: {num_batches}")
    print(f"  - 每个路径的训练轮数: {num_rounds}")
    print(f"  - 每个样本的步数: {EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(
        f"  - 总训练循环: {num_paths} 个路径 x {num_rounds} 轮 x {num_batches} 批 = {num_paths * num_rounds * num_batches} 批次")
    print("-" * 80)

    # 逐个路径训练:完成一个路径的所有轮次后,再训练下一个路径
    for path_idx in range(num_paths):
        target_path = target_paths[path_idx]
        print(f"\n{'=' * 80}")
        print(f"开始训练路径 {path_idx + 1}/{num_paths}")
        print(f"目标路径: {sorted(target_path)}")
        print(f"使用缓冲区: replay_buffers[{path_idx}]")
        print(f"{'=' * 80}")

        # 每个路径训练NUM_ROUNDS轮
        for round_idx in range(num_rounds):
            print(f"\n{'*' * 80}")
            print(f"路径 {path_idx + 1} - 第 {round_idx + 1}/{num_rounds} 轮训练")
            print(f"{'*' * 80}")

            # 每轮遍历num_batches个批次
            for batch_idx in range(num_batches):
                print(f"\n  批次 {batch_idx + 1}/{num_batches} (路径 {path_idx + 1}, 轮次 {round_idx + 1})")

                # 获取当前批次的样本
                batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]

                batch_rewards = []
                batch_similarities = []

                # 对批次中的每个样本进行训练
                for sample_idx, initial_state in enumerate(batch_samples):
                    state = initial_state.copy()
                    episode_reward = 0
                    final_similarity = 0

                    # 每个样本执行STEPS_PER_SAMPLE步
                    for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                        action_delta, action_idx = agent.get_action(state)

                        next_state = state + action_delta
                        next_state = clip_state(next_state)

                        triggered = execute_Tr(*next_state)  # 执行测试需求
                        reward = unified_reward_function(triggered, target_path)
                        similarity = coverage_similarity(triggered, target_path)

                        done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)

                        # 存储经验到对应路径的缓冲区
                        agent.store_experience(
                            path_idx, state, action_idx, reward, next_state, done, similarity
                        )

                        state = next_state
                        episode_reward += reward
                        final_similarity = similarity
                        total_steps += 1

                    batch_rewards.append(episode_reward)
                    batch_similarities.append(final_similarity)

                # 批次统计
                avg_reward = np.mean(batch_rewards)
                avg_similarity = np.mean(batch_similarities)
                max_similarity = np.max(batch_similarities)
                print(f"    平均奖励={avg_reward:.2f}, 平均相似度={avg_similarity:.4f}, "
                      f"最大相似度={max_similarity:.4f}, epsilon={agent.epsilon:.3f}")

                # 从对应路径的缓冲区进行训练
                print(f"    从缓冲区训练 (路径 {path_idx})...")
                agent.replay_train(path_idx)

                # 显示当前路径的缓冲区大小
                buffer_size = len(agent.replay_buffers[path_idx])
                print(f"    路径 {path_idx} 缓冲区大小: {buffer_size}, 总训练次数: {agent.replay_train_count}")

    training_time = time.time() - start_time

    print("\n" + "=" * 80)
    print(f"DQN训练完成! 总耗时: {training_time:.2f} 秒, 总步数: {total_steps}")
    print(f"总训练次数: {agent.replay_train_count}")
    print(f"目标网络更新次数: {agent.replay_train_count // 2}")

    # 显示每个路径的缓冲区大小
    print("\n各路径缓冲区大小:")
    buffer_stats = agent.get_buffer_stats()
    for path_idx, size in buffer_stats.items():
        print(f"  路径 {path_idx + 1}: {size} 个经验")

    print("=" * 80)

    # 获取Top-K结果
    print(f"\n提取每个路径相似度最高的{EXPERIMENT_CONFIG['TOP_K_SAMPLES']}个样本...")
    dqn_top_k_results = agent.get_all_top_k(EXPERIMENT_CONFIG['TOP_K_SAMPLES'])

    return agent, dqn_top_k_results, training_time, total_steps, agent.replay_train_count


# === 主函数 ===
def main():
    print("\n" + "=" * 80)
    print("DQN - 20次运行实验")
    print("性能指标计算")
    print("=" * 80)

    all_dqn_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    # 执行20次运行
    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'='*80}")
        print(f"开始第 {run_idx + 1}/{EXPERIMENT_CONFIG['NUM_RUNS']} 次运行")
        print(f"{'='*80}")

        # 执行DQN训练
        dqn_agent, dqn_results, training_time, total_steps, update_count = train_dqn_workflow()

        # 计算性能指标
        performance_data = calculate_run_performance(
            run_idx, dqn_results, training_time, total_steps, update_count, dqn_agent
        )

        # 存储结果
        all_dqn_results.append(dqn_results)
        all_performance_data.append(performance_data)

        print(f"\n第 {run_idx + 1} 次运行完成!")
        print(f"  总奖励: {performance_data['Total Reward']}")
        print(f"  平均奖励: {performance_data['Average Reward']}")
        print(f"  收敛性: {performance_data['Convergence']}")

    # 导出Excel(包含20次运行的所有结果)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"DQN_20次运行_{timestamp}.xlsx"
    export_to_excel(all_dqn_results, all_performance_data, target_paths, output_path)

    # 输出汇总统计
    print("\n" + "=" * 80)
    print("20次运行汇总统计")
    print("=" * 80)

    # 计算各项指标的均值和标准差
    # 提取性能指标
    total_rewards = [p['Total Reward'] for p in all_performance_data]
    average_rewards = [p['Average Reward'] for p in all_performance_data]
    convergences = [p['Convergence'] for p in all_performance_data]
    environment_adaptabilities = [p['Environment Adaptability'] for p in all_performance_data]
    generalization_abilities = [p['Generalization Ability'] for p in all_performance_data]
    computational_efficiencies = [p['Computational Efficiency'] for p in all_performance_data]
    policy_update_frequencies = [p['Policy Update Frequency'] for p in all_performance_data]
    avg_similarities = [p['Average Similarity'] for p in all_performance_data]

    print(f"\n总奖励统计:")
    print(f"  均值: {np.mean(total_rewards):.2f}")
    print(f"  标准差: {np.std(total_rewards):.2f}")

    print(f"\n平均奖励统计:")
    print(f"  均值: {np.mean(average_rewards):.4f}")
    print(f"  标准差: {np.std(average_rewards):.4f}")

    print(f"\n收敛性统计:")
    print(f"  均值: {np.mean(convergences):.4f}")
    print(f"  标准差: {np.std(convergences):.4f}")

    print(f"\n环境适应性统计:")
    print(f"  均值: {np.mean(environment_adaptabilities):.4f}")
    print(f"  标准差: {np.std(environment_adaptabilities):.4f}")

    print(f"\n泛化能力统计:")
    print(f"  均值: {np.mean(generalization_abilities):.4f}")
    print(f"  标准差: {np.std(generalization_abilities):.4f}")

    print(f"\n计算效率统计:")
    print(f"  均值: {np.mean(computational_efficiencies):.2f}")
    print(f"  标准差: {np.std(computational_efficiencies):.2f}")

    print(f"\n策略更新频率统计:")
    print(f"  均值: {np.mean(policy_update_frequencies):.4f}")
    print(f"  标准差: {np.std(policy_update_frequencies):.4f}")

    print(f"\n平均相似度统计:")
    print(f"  均值: {np.mean(avg_similarities):.4f}")
    print(f"  标准差: {np.std(avg_similarities):.4f}")

    print("\n" + "=" * 80)
    print(f"所有 {EXPERIMENT_CONFIG['NUM_RUNS']} 次运行完成!")
    print("=" * 80)


if __name__ == "__main__":
    main()