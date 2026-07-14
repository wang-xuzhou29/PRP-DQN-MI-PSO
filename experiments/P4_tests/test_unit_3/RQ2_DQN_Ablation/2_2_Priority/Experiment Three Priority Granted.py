import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# === 设备设置 ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === 状态空间范围 ===
X_MIN = 1
X_MAX = 200
Y_MIN = 1
Y_MAX = 200
Z_MIN = 2
Z_MAX = 150


# === 状态归一化/反归一化函数 ===
def normalize_state(state):
    """将状态归一化到[0,1]范围"""
    x, y, z = state
    norm_x = (x - X_MIN) / (X_MAX - X_MIN)
    norm_y = (y - Y_MIN) / (Y_MAX - Y_MIN)
    norm_z = (z - Z_MIN) / (Z_MAX - Z_MIN)
    return (norm_x, norm_y, norm_z)


def denormalize_state(norm_state):
    """将归一化状态还原为原始范围"""
    norm_x, norm_y, norm_z = norm_state
    x = int(round(norm_x * (X_MAX - X_MIN) + X_MIN))
    y = int(round(norm_y * (Y_MAX - Y_MIN) + Y_MIN))
    z = int(round(norm_z * (Z_MAX - Z_MIN) + Z_MIN))
    x = max(X_MIN, min(X_MAX, x))
    y = max(Y_MIN, min(Y_MAX, y))
    z = max(Z_MIN, min(Z_MAX, z))
    return (x, y, z)


def is_valid_state(state):
    """检查状态是否在有效范围内"""
    x, y, z = state
    return (X_MIN <= x <= X_MAX and
            Y_MIN <= y <= Y_MAX and
            Z_MIN <= z <= Z_MAX)


def clip_state(state):
    """将状态裁剪到有效范围内"""
    x, y, z = state
    return (
        max(X_MIN, min(X_MAX, x)),
        max(Y_MIN, min(Y_MAX, y)),
        max(Z_MIN, min(Z_MAX, z))
    )


# === 优先级指标收集器 ===
class PrioritizedMetricsCollector:
    def __init__(self, experiment_name="Prioritized_DQN"):
        self.experiment_name = experiment_name
        self.start_time = None
        self.end_time = None
        self.total_reward = 0
        self.td_errors = []
        self.final_output_similarities = []
        self.action_improvements = []
        self.total_memory_usage = 0
        self.memory_check_count = 0
        self.step_count = 0
        self.episode_rewards = []
        self.episode_similarities = []
        self.episode_td_errors = []
        self.episode_epsilon_values = []
        self.episode_memory_usage = []
        self.priority_statistics = []
        self.importance_weights = []
        self.high_priority_samples_ratio = []
        self.priority_distribution_stats = []
        self.similar_paths_performance = []
        self.isolated_paths_performance = []
        self.milestone_data = {}
        self.convergence_window = 20
        self.convergence_threshold = 0.02
        self.convergence_detected_episode = None
        self.sample_efficiency_data = []
        self.performance_milestones = [0.6, 0.7, 0.75, 0.8]
        self.learning_curve_characteristics = {}
        self.early_vs_late_performance = {}

    def reset(self):
        self.start_time = None
        self.end_time = None
        self.total_reward = 0
        self.td_errors = []
        self.final_output_similarities = []
        self.action_improvements = []
        self.total_memory_usage = 0
        self.memory_check_count = 0
        self.step_count = 0
        self.episode_rewards = []
        self.episode_similarities = []
        self.episode_td_errors = []
        self.episode_epsilon_values = []
        self.episode_memory_usage = []
        self.priority_statistics = []
        self.importance_weights = []
        self.high_priority_samples_ratio = []
        self.priority_distribution_stats = []
        self.similar_paths_performance = []
        self.isolated_paths_performance = []
        self.milestone_data = {}
        self.convergence_detected_episode = None
        self.sample_efficiency_data = []
        self.learning_curve_characteristics = {}
        self.early_vs_late_performance = {}

    def start_training(self):
        self.start_time = time.time()

    def end_training(self):
        self.end_time = time.time()

    def record_step_metrics(self, reward, td_error, triggered, target_path, priority=None, is_weight=None):
        self.step_count += 1
        self.total_reward += reward
        self.td_errors.append(td_error)
        if priority is not None:
            self.priority_statistics.append(priority)
        if is_weight is not None:
            self.importance_weights.append(is_weight)
        process = psutil.Process(os.getpid())
        current_memory = process.memory_info().rss / 1024 / 1024
        self.total_memory_usage += current_memory
        self.memory_check_count += 1

    def record_episode_metrics(self, episode, episode_reward, avg_similarity, avg_td_error, epsilon,
                               path_group="similar", priority_stats=None):
        self.episode_rewards.append(episode_reward)
        self.episode_similarities.append(avg_similarity)
        self.episode_td_errors.append(avg_td_error)
        self.episode_epsilon_values.append(epsilon)
        process = psutil.Process(os.getpid())
        current_memory = process.memory_info().rss / 1024 / 1024
        self.episode_memory_usage.append(current_memory)
        if priority_stats:
            self.priority_distribution_stats.append({
                'episode': episode,
                'mean_priority': priority_stats.get('mean_priority', 0),
                'max_priority': priority_stats.get('max_priority', 0),
                'min_priority': priority_stats.get('min_priority', 0),
                'high_priority_ratio': priority_stats.get('high_priority_ratio', 0)
            })
        if path_group == "similar":
            self.similar_paths_performance.append({
                'episode': episode, 'reward': episode_reward,
                'similarity': avg_similarity, 'td_error': avg_td_error
            })
        else:
            self.isolated_paths_performance.append({
                'episode': episode, 'reward': episode_reward,
                'similarity': avg_similarity, 'td_error': avg_td_error
            })
        if episode in [50, 100, 150, 200, 250, 300, 400, 450, 500]:
            self.milestone_data[episode] = {
                'avg_reward': np.mean(self.episode_rewards[-10:]) if len(self.episode_rewards) >= 10 else episode_reward,
                'avg_similarity': avg_similarity,
                'avg_td_error': avg_td_error,
                'epsilon': epsilon,
                'memory_usage': current_memory,
                'total_steps': self.step_count,
                'priority_stats': priority_stats
            }
        self._check_convergence(episode)
        self._check_performance_milestones(episode, avg_similarity)

    def _check_convergence(self, episode):
        if len(self.episode_similarities) >= self.convergence_window and self.convergence_detected_episode is None:
            recent_similarities = self.episode_similarities[-self.convergence_window:]
            if np.std(recent_similarities) < self.convergence_threshold:
                self.convergence_detected_episode = episode

    def _check_performance_milestones(self, episode, similarity):
        for milestone in self.performance_milestones:
            if similarity >= milestone and not any(data[1] == milestone for data in self.sample_efficiency_data):
                self.sample_efficiency_data.append((episode, milestone, self.step_count))

    def record_final_output_sample(self, triggered, target_path):
        if len(triggered | target_path) > 0:
            similarity = len(triggered & target_path) / len(triggered | target_path)
        else:
            similarity = 0.0
        self.final_output_similarities.append(similarity)

    def record_action_improvement(self, current_reward, prev_reward):
        if prev_reward is not None:
            improvement = current_reward - prev_reward
            self.action_improvements.append(1 if improvement > 0 else 0)


# 全局指标收集器
prioritized_metrics = PrioritizedMetricsCollector("Prioritized_DQN_Enhanced")


# === 奖励函数 ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward


# === 验证规则执行函数 ===
def execute_Tr(x, y, z):
    """执行验证规则，返回触发的规则集合"""
    b = set()

    # 异常类型1：质量参数乘积异常
    if ((y * z) / (x + 1) > 80) != ((y * y * z) / (x + 1) > 80):
        b.add(1)
    if ((y * z) / (x + 1) > 80) != ((y * z * z) / (x + 1) > 80):
        b.add(2)
    if ((y * z) / (x + 1) > 80) != ((y * x * z) / (x + 1) > 80):
        b.add(3)
    if ((y * z) / (x + 1) > 80) != ((y * z) / (x + 1) > 60):
        b.add(4)
    if ((y * z) / (x + 1) > 80) != ((y * z) / (x + 10) > 80):
        b.add(5)
    if ((y * z) / (x + 1) > 80) != ((y * z) / (x + 13) > 80):
        b.add(6)
    if ((y * z) / (x + 1) > 80) != ((y * z * 5) / (x + 1) > 80):
        b.add(7)
    if ((y * z) / (x + 1) > 80) != ((y * z * 2) / (x + 1) > 80):
        b.add(8)
    if ((y * z) / (x + 1) > 80) != ((y * z) / (x + 1) > 40):
        b.add(9)
    if ((y * z) / (x + 1) > 80) != ((y * x) / (x + 1) > 80):
        b.add(10)
    if ((y * z) / (x + 1) > 80) != ((y * y) / (x + 1) > 80):
        b.add(11)
    if ((y * z) / (x + 1) > 80) != ((z * z) / (x + 1) > 80):
        b.add(12)

    # 异常类型2：质量差值异常
    if ((z - x) < 0.4 * y) != ((z - x) < 0.3 * y):
        b.add(13)
    if ((z - x) < 0.4 * y) != ((z - x) < 0.5 * y):
        b.add(14)
    if ((z - x) < 0.4 * y) != ((z - x) < 0.4 * z):
        b.add(15)
    if ((z - x) < 0.4 * y) != ((z - x) < 0.4 * x):
        b.add(16)
    if ((z - x) < 0.4 * y) != ((z * 1.1 - x) < 0.4 * y):
        b.add(17)
    if ((z - x) < 0.4 * y) != ((z * 2 - x) < 0.4 * y):
        b.add(18)
    if ((z - x) < 0.4 * y) != ((z * z - x) < 0.4 * y):
        b.add(19)
    if ((z - x) < 0.4 * y) != ((z * x - x) < 0.4 * y):
        b.add(20)
    if ((z - x) < 0.4 * y) != ((z * y - x) < 0.4 * y):
        b.add(21)
    if ((z - x) < 0.4 * y) != ((z * 1.5 - x) < 0.4 * y):
        b.add(22)

    # 异常类型3：质量立方关系
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2 + y ** 3) < z ** 2):
        b.add(23)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 2) < z ** 2):
        b.add(24)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 1):
        b.add(25)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 3):
        b.add(26)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 4):
        b.add(27)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 4) < z ** 2):
        b.add(28)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + x ** 3) < z ** 2):
        b.add(29)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 1 + y ** 3) < z ** 2):
        b.add(30)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 1) < z ** 2):
        b.add(31)
    if ((x ** 3 + y ** 3) < z ** 2) != (((x ** 3) * 2 + y ** 3) < z ** 2):
        b.add(32)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + (y ** 3) * 2) < z ** 2):
        b.add(33)
    if ((x ** 3 + y ** 3) < z ** 2) != ((y ** 3 + y ** 3) < z ** 2):
        b.add(34)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < (z ** 2) * 2):
        b.add(35)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < (x ** 2) * 2):
        b.add(36)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < (y ** 2) * 2):
        b.add(37)

    # 异常类型6：质量同步性检查
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 2 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1):
        b.add(38)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 3 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1):
        b.add(39)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 2) < 0.1 and abs(y % 1 - z % 1) < 0.1):
        b.add(40)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 3) < 0.1 and abs(y % 1 - z % 1) < 0.1):
        b.add(41)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 5) < 0.1 and abs(y % 1 - z % 1) < 0.1):
        b.add(42)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 5 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1):
        b.add(43)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 1) < 0.1 and abs(y % 2 - z % 1) < 0.1):
        b.add(44)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 1) < 0.1 and abs(y % 3 - z % 1) < 0.1):
        b.add(45)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 1) < 0.1 and abs(y % 5 - z % 1) < 0.1):
        b.add(46)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 2) < 0.1):
        b.add(47)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 3) < 0.1):
        b.add(48)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 5) < 0.1):
        b.add(49)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 1) < 0.1 and abs(y % 4 - z % 1) < 0.1):
        b.add(50)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 4) < 0.1):
        b.add(51)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 4 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1):
        b.add(52)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 6) < 0.1 and abs(y % 1 - z % 1) < 0.1):
        b.add(53)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 6 - y % 1) < 0.1 and abs((y * 2) % 1 - z % 1) < 0.1):
        b.add(54)
    if (abs(x % 1 - y % 1) < 0.1 and abs(y % 1 - z % 1) < 0.1) != (
            abs(x % 1 - y % 1) < 0.1 and abs(y % 6 - (z * 2) % 1) < 0.1):
        b.add(55)

    # 其他复杂检查逻辑
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * y * z > 500000 and (x + y + z) / 2 < 85):
        b.add(56)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * y * z > 500000 and (x + y + z) / 4 < 85):
        b.add(57)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * y * z > 500000 and (x * 2 + y + z) / 3 < 85):
        b.add(58)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * y * z > 500000 and (x + y * 2 + z) / 3 < 85):
        b.add(59)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * y * z > 500000 and (x + y - z) / 3 < 85):
        b.add(60)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * y * z > 500000 and (x + y * z * 2) / 3 < 85):
        b.add(61)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * y * z > 500000 and (x + y + x) / 3 < 85):
        b.add(62)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * y * z > 500000 and (x + y + y) / 3 < 85):
        b.add(63)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * y * z > 600000 and (x + y + z) / 3 < 85):
        b.add(64)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * y * z * 2 > 500000 and (x + y + z) / 3 < 85):
        b.add(65)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            y * y * z > 500000 and (x + y + z) / 3 < 85):
        b.add(66)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            z * y * z > 500000 and (x + y + z) / 3 < 85):
        b.add(67)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * x * z > 500000 and (x + y + z) / 3 < 85):
        b.add(68)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * y * x > 500000 and (x + y + z) / 3 < 85):
        b.add(69)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * z * z > 500000 and (x + y + z) / 3 < 85):
        b.add(70)
    if (x * y * z > 500000 and (x + y + z) / 3 < 85) != (
            x * y * y > 500000 and (x + y + z) / 3 < 85):
        b.add(71)

    return b


def jaccard_similarity(set1, set2):
    """计算Jaccard相似度"""
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0


# === 目标路径定义 ===
targetPaths = [
    {1, 2, 3, 4, 7, 8, 9, 10, 11, 19, 20, 21, 27, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 49, 50, 51, 52, 53, 54, 55,
     56, 58, 59, 61, 62, 63, 64, 67, 68, 70},
    {1, 2, 3, 4, 7, 8, 9, 12, 18, 19, 20, 21, 22, 27, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53,
     54, 55, 56, 58, 59, 61, 64, 66, 71},
    {1, 2, 3, 4, 7, 8, 9, 12, 13, 17, 18, 19, 20, 21, 22, 26, 27, 40, 41, 42, 44, 45, 46, 48, 49, 50, 51, 52, 53, 55,
     56, 58, 59, 61, 64, 66, 69, 71},
    {5, 6, 10, 11, 13, 17, 18, 19, 20, 21, 22, 26, 27, 38, 39, 40, 42, 43, 44, 46, 49, 50, 52, 53, 54, 55, 56, 58, 59,
     61, 64, 66, 69, 71},
    {1, 2, 3, 7, 12, 17, 18, 19, 20, 21, 22, 26, 27, 30, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 49, 50, 51, 52, 53, 54,
     55, 65, 68, 70},
    {1, 2, 3, 7, 8, 9, 12, 17, 18, 19, 20, 21, 22, 26, 27, 38, 39, 40, 41, 42, 44, 45, 46, 47, 48, 50, 51, 52, 53, 54,
     55, 57, 60, 63},
    {1, 2, 3, 4, 7, 8, 9, 12, 23, 24, 26, 27, 30, 31, 34, 35, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51,
     52, 53, 54, 55},
    {16, 18, 19, 20, 21, 22, 27, 38, 40, 41, 42, 43, 44, 45, 46, 48, 49, 50, 51, 52, 53, 54, 55, 56, 58, 59, 61, 63, 64,
     68, 69, 70},
    {1, 2, 3, 4, 7, 8, 9, 12, 14, 15, 16, 26, 27, 38, 40, 41, 42, 43, 44, 45, 46, 48, 49, 50, 51, 52, 53, 54, 55, 57,
     60, 62, 63},
    {5, 6, 10, 13, 15, 16, 18, 19, 20, 21, 22, 27, 29, 31, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53,
     54, 55},
    {1, 2, 3, 7, 8, 9, 12, 25, 28, 29, 32, 33, 36, 37, 38, 40, 41, 42, 43, 44, 45, 46, 47, 48, 50, 51, 52, 53, 54, 55,
     70}
]


# === 路径相似度计算和分组 ===
def compute_path_similarity_matrix(paths):
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix


def group_paths_by_similarity(paths):
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.mean(avg_sim_scores)
    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]
    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)
    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group


# === 优先经验回放 ===
class PrioritizedExperienceReplay:
    def __init__(self, capacity=10000, alpha=0.6, beta_start=0.4, beta_frames=100000):
        self.capacity = capacity
        self.alpha = alpha
        self.beta_start = beta_start
        self.beta_frames = beta_frames
        self.frame = 1
        self.buffer = []
        self.priorities = np.zeros(capacity, dtype=np.float32)
        self.pos = 0
        self.size = 0
        self.max_priority = 1.0
        self.min_priority = 1.0

    def beta(self):
        return min(1.0, self.beta_start + (1.0 - self.beta_start) * self.frame / self.beta_frames)

    def append(self, experience):
        if len(self.buffer) < self.capacity:
            self.buffer.append(experience)
        else:
            self.buffer[self.pos] = experience
        self.priorities[self.pos] = self.max_priority
        self.pos = (self.pos + 1) % self.capacity
        self.size = min(self.size + 1, self.capacity)

    def sample(self, batch_size):
        if self.size < batch_size:
            return [], [], []
        priorities = self.priorities[:self.size]
        probs = priorities ** self.alpha
        probs /= probs.sum()
        indices = np.random.choice(self.size, batch_size, p=probs, replace=False)
        unique_batch = []
        unique_indices = []
        seen_states = set()
        for idx in indices:
            experience = self.buffer[idx]
            state_tensor = experience[0]
            state_tuple = tuple(state_tensor.cpu().numpy().flatten())
            if state_tuple not in seen_states:
                seen_states.add(state_tuple)
                unique_batch.append(experience)
                unique_indices.append(idx)
        if len(unique_batch) < batch_size:
            remaining_indices = [i for i in range(self.size) if i not in unique_indices]
            if remaining_indices:
                remaining_probs = priorities[remaining_indices] ** self.alpha
                remaining_probs /= remaining_probs.sum()
                needed = batch_size - len(unique_batch)
                additional_indices = np.random.choice(remaining_indices, min(needed, len(remaining_indices)),
                                                      p=remaining_probs, replace=False)
                for idx in additional_indices:
                    experience = self.buffer[idx]
                    state_tensor = experience[0]
                    state_tuple = tuple(state_tensor.cpu().numpy().flatten())
                    if state_tuple not in seen_states:
                        seen_states.add(state_tuple)
                        unique_batch.append(experience)
                        unique_indices.append(idx)
        total = len(self.buffer)
        unique_indices = np.array(unique_indices)
        weights = (total * probs[unique_indices]) ** (-self.beta())
        weights /= weights.max()
        self.frame += 1
        return unique_batch, unique_indices, weights

    def update_priorities(self, indices, priorities):
        for idx, priority in zip(indices, priorities):
            if idx < self.size:
                self.priorities[idx] = priority
                self.max_priority = max(self.max_priority, priority)
                self.min_priority = min(self.min_priority, priority)

    def get_priority_statistics(self):
        if self.size == 0:
            return None
        priorities = self.priorities[:self.size]
        mean_priority = np.mean(priorities)
        max_priority = np.max(priorities)
        min_priority = np.min(priorities)
        high_priority_ratio = np.mean(priorities > mean_priority)
        return {
            'mean_priority': mean_priority,
            'max_priority': max_priority,
            'min_priority': min_priority,
            'high_priority_ratio': high_priority_ratio
        }

    def __len__(self):
        return self.size

    def get_high_reward_samples(self, target_path, num_samples=20):
        if len(self.buffer) == 0:
            return []
        samples_with_recalculated_scores = []
        for experience in self.buffer:
            norm_state_tensor = experience[0]
            norm_state = tuple(norm_state_tensor.cpu().numpy().flatten())
            state_tuple = denormalize_state(norm_state)
            triggered = execute_Tr(state_tuple[0], state_tuple[1], state_tuple[2])
            new_reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_recalculated_scores.append((state_tuple, new_reward, sim, triggered))
        samples_with_recalculated_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_recalculated_scores[:num_samples]


def load_path_data(file_path):
    path_data = []
    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines[2:]:
            parts = line.strip().split("\t")
            state = tuple(map(int, parts[0].split()))
            path_data.append(state)
    return path_data


# === DQN网络 ===
class DQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, 128)
        self.fc2 = nn.Linear(128, 64)
        self.fc3 = nn.Linear(64, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)


# === 优先级DQN智能体 ===
class PrioritizedDQNAgent:
    def __init__(self, state_dim, action_dim, replay_buffer, gamma=0.99, epsilon=1.0,
                 epsilon_decay=0.995, epsilon_min=0.1, learning_rate=0.001):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer
        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        delta_values = [35, 25, 10, 5, 2, -2, -5, -10, -25, -35]
        dim = action_idx // 10
        delta_idx = action_idx % 10
        delta = delta_values[delta_idx]
        if dim == 0:
            return (delta, 0, 0)
        elif dim == 1:
            return (0, delta, 0)
        elif dim == 2:
            return (0, 0, delta)
        return (0, 0, 0)

    def act(self, state):
        if random.random() < self.epsilon:
            return random.randrange(self.action_dim)
        normalized_state = normalize_state(state)
        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)
        return torch.argmax(q_values, dim=1).item()

    def store_transition(self, state, action, reward, next_state, done):
        normalized_state = normalize_state(state)
        normalized_next_state = normalize_state(next_state)
        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        next_state_tensor = torch.tensor(normalized_next_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)
            next_q_values = self.target_model(next_state_tensor)
            max_next_q_values = next_q_values.max(1)[0]
            target_q_values = reward + (self.gamma * max_next_q_values * (1 - done))
            td_error = torch.abs(q_values[0][action] - target_q_values).item()
        self.replay_buffer.append((state_tensor, action, reward, next_state_tensor, done, td_error))
        return td_error

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return
        batch, batch_indices, is_weights = self.replay_buffer.sample(batch_size)
        if not batch:
            return
        states, actions, rewards, next_states, dones, _ = zip(*batch)
        states = torch.tensor(np.array([s.cpu().numpy().flatten() for s in states]), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array([ns.cpu().numpy().flatten() for ns in next_states]), dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)
        is_weights = torch.tensor(is_weights, dtype=torch.float32).to(device)
        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))
        td_errors = torch.abs(current_q_values - target_q_values)
        loss = (is_weights * (current_q_values - target_q_values) ** 2).mean()
        self.optimizer.zero_grad()
        loss.backward()
        self.optimizer.step()
        new_priorities = td_errors.detach().cpu().numpy() + 1e-6
        self.replay_buffer.update_priorities(batch_indices, new_priorities)
        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())


# === 相似路径样本生成 ===
def generate_samples_for_similar_paths(similar_group_indices, num_total=2000, top_k=200):
    def jaccard_similarity_local(a, b):
        if not a and not b: return 1.0
        return len(a & b) / len(a | b) if a | b else 0.0

    def compute_robustness(state, path):
        base = execute_Tr(*state)
        if not base: return 0.0
        rob, neighbors = 0.0, 0
        for dx in [-1, 0, 1]:
            for dy in [-1, 0, 1]:
                for dz in [-1, 0, 1]:
                    if dx == dy == dz == 0: continue
                    neighbor_state = (state[0] + dx, state[1] + dy, state[2] + dz)
                    if not is_valid_state(neighbor_state): continue
                    neighbor = clip_state(neighbor_state)
                    n_trig = execute_Tr(*neighbor)
                    if not n_trig: continue
                    rob += jaccard_similarity_local(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"prioritized_path{path_id}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Prioritized Path {path_id}\n")
            f.write("x y z\tScore\tSimilarity\tLengthDiff\tRobustness\n")
            for s in samples:
                x, y, z = s[0]
                f.write(f"{x} {y} {z}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_prioritized"
    for path_idx in similar_group_indices:
        path = targetPaths[path_idx]
        samples = []
        attempts = 0
        while len(samples) < top_k and attempts < num_total * 5:
            attempts += 1
            state = (
                np.random.randint(X_MIN, X_MAX + 1),
                np.random.randint(Y_MIN, Y_MAX + 1),
                np.random.randint(Z_MIN, Z_MAX + 1)
            )
            triggered = execute_Tr(*state)
            if not triggered: continue
            sim = jaccard_similarity_local(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            score = 0.55 * sim + 0.25 * len_diff + 0.2 * rob
            samples.append((state, score, sim, len_diff, rob))
        if samples:
            samples.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples[:top_k], base_dir=base_dir)


# === 相似路径训练 ===
def prioritized_generate_and_train_for_similar_paths(agent, similar_group, path_documents, episodes=500,
                                                     batch_size=32, is_isolated=False):
    trained_paths = set()
    global_replay_count = 0

    for episode in range(episodes):
        episode_reward = 0
        episode_similarities = []
        episode_td_errors = []

        for path_idx in similar_group:
            if path_idx in trained_paths:
                continue
            file_path = os.path.join(path_documents,
                                     f"prioritized_path{path_idx + 1}{'_isolated' if is_isolated else ''}.txt")
            path_data = load_path_data(file_path)
            target_path = targetPaths[path_idx]

            BATCH_SIZE = 50
            N_SAMPLES = 200
            N_STEPS = 10
            REPLAY_TIMES = 3

            for batch_start in range(0, N_SAMPLES, BATCH_SIZE):
                batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)
                for test_data_idx in range(batch_start, batch_end):
                    if test_data_idx >= len(path_data):
                        break
                    state = path_data[test_data_idx]
                    prev_state = None
                    prev_triggered = None
                    prev_reward = None

                    for step in range(N_STEPS):
                        legal_actions = []
                        for a in range(agent.action_dim):
                            dx, dy, dz = agent.decode_action(a)
                            cand_next = (state[0] + dx, state[1] + dy, state[2] + dz)
                            if is_valid_state(cand_next):
                                legal_actions.append(a)
                        if not legal_actions:
                            break
                        if random.random() < agent.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            normalized_state = normalize_state(state)
                            state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        dx, dy, dz = agent.decode_action(action)
                        next_state = clip_state((state[0] + dx, state[1] + dy, state[2] + dz))
                        triggered = execute_Tr(*next_state)
                        reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                        done = (step == N_STEPS - 1)
                        td_error = agent.store_transition(state, action, reward, next_state, done)

                        priority_stats = agent.replay_buffer.get_priority_statistics()
                        current_priority = priority_stats['mean_priority'] if priority_stats else 0
                        prioritized_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                current_priority, None)
                        episode_similarities.append(jaccard_similarity(triggered, target_path))
                        episode_td_errors.append(td_error)
                        if prev_reward is not None:
                            prioritized_metrics.record_action_improvement(reward, prev_reward)
                        prev_state = state
                        prev_triggered = triggered
                        prev_reward = reward
                        state = next_state
                        episode_reward += reward

                for _ in range(REPLAY_TIMES):
                    if len(agent.replay_buffer) >= batch_size:
                        agent.train(batch_size)
                        global_replay_count += 1
                        if global_replay_count % 2 == 0:
                            agent.update_target_model()
            trained_paths.add(path_idx)

        avg_similarity = np.mean(episode_similarities) if episode_similarities else 0
        avg_td_error = np.mean(episode_td_errors) if episode_td_errors else 0
        priority_stats = agent.replay_buffer.get_priority_statistics()
        prioritized_metrics.record_episode_metrics(episode, episode_reward, avg_similarity, avg_td_error,
                                                   agent.epsilon, "similar", priority_stats)
        if len(trained_paths) == len(similar_group):
            break
    return agent


def generate_samples_for_isolated_paths_prioritized(agent_similar, isolated_group_indices, num_total=2000, top_k=200):
    def compute_q_value(state, agent):
        normalized_state = normalize_state(state)
        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = agent.model(state_tensor)
        return q_values.max().item()

    def compute_robustness(state, path):
        base = execute_Tr(*state)
        if not base: return 0.0
        rob, neighbors = 0.0, 0
        for dx in [-1, 0, 1]:
            for dy in [-1, 0, 1]:
                for dz in [-1, 0, 1]:
                    if dx == dy == dz == 0: continue
                    neighbor_state = (state[0] + dx, state[1] + dy, state[2] + dz)
                    if not is_valid_state(neighbor_state): continue
                    neighbor = clip_state(neighbor_state)
                    n_trig = execute_Tr(*neighbor)
                    if not n_trig: continue
                    rob += jaccard_similarity(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"prioritized_path{path_id}_isolated.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Prioritized Isolated Path {path_id}\n")
            f.write("x y z\tScore\tSimilarity\tLengthDiff\tRobustness\tQ_value_complement\n")
            for s in samples:
                x, y, z = s[0]
                f.write(f"{x} {y} {z}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\t{s[5]:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_prioritized"
    for path_idx in isolated_group_indices:
        path = targetPaths[path_idx]
        samples_raw = []
        attempts = 0
        print(f"生成隔离路径 {path_idx + 1} 样本...")
        while len(samples_raw) < num_total and attempts < num_total * 5:
            attempts += 1
            state = (
                np.random.randint(X_MIN, X_MAX + 1),
                np.random.randint(Y_MIN, Y_MAX + 1),
                np.random.randint(Z_MIN, Z_MAX + 1)
            )
            triggered = execute_Tr(*state)
            if not triggered: continue
            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            q_value = compute_q_value(state, agent_similar)
            samples_raw.append((state, sim, len_diff, rob, q_value))

        if not samples_raw:
            print(f"路径 {path_idx + 1}: 未生成有效样本")
            continue

        q_values_list = [s[4] for s in samples_raw]
        q_min = min(q_values_list)
        q_max = max(q_values_list)
        print(f"路径 {path_idx + 1}: 生成了 {len(samples_raw)} 个候选样本, Q值范围: [{q_min:.4f}, {q_max:.4f}]")

        samples_final = []
        for state, sim, len_diff, rob, q_value in samples_raw:
            if q_max - q_min > 1e-6:
                q_normalized = (q_value - q_min) / (q_max - q_min)
            else:
                q_normalized = 0.5
            q_complement = 1.0 - q_normalized
            score = 0.28 * sim + 0.1 * len_diff + 0.19 * rob + 0.43 * q_complement
            samples_final.append((state, score, sim, len_diff, rob, q_complement))

        if samples_final:
            samples_final.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples_final[:top_k], base_dir=base_dir)
            print(f"路径 {path_idx + 1}: 保存了 {min(top_k, len(samples_final))} 个最终样本")


# === 隔离路径训练 ===
def prioritized_generate_and_train_for_isolated_paths(agent_similar, agent_isolated, similar_group,
                                                      isolated_group, path_documents, episodes=500, batch_size=32,
                                                      is_isolated=True):
    trained_paths = set()
    global_replay_count = 0
    stage1_samples_pool = {}

    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_similar.replay_buffer.get_high_reward_samples(target_path, num_samples=100)
        stage1_samples_pool[path_idx] = high_reward_samples

    for episode in range(episodes):
        episode_reward = 0
        episode_similarities = []
        episode_td_errors = []

        for path_idx in isolated_group:
            if path_idx in trained_paths:
                continue
            file_path = os.path.join(path_documents, f"prioritized_path{path_idx + 1}_isolated.txt")
            stage2_path_data = load_path_data(file_path)
            stage1_samples = stage1_samples_pool.get(path_idx, [])
            target_path = targetPaths[path_idx]

            BATCH_SIZE = 50
            N_SAMPLES_STAGE2 = min(140, len(stage2_path_data))
            N_SAMPLES_STAGE1 = min(60, len(stage1_samples))
            N_STEPS = 10
            REPLAY_TIMES = 3

            for batch_start in range(0, N_SAMPLES_STAGE2, BATCH_SIZE):
                batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES_STAGE2)
                for test_data_idx in range(batch_start, batch_end):
                    if test_data_idx >= len(stage2_path_data):
                        break
                    state = stage2_path_data[test_data_idx]
                    prev_state = None
                    prev_triggered = None
                    prev_reward = None

                    for step in range(N_STEPS):
                        legal_actions = []
                        for a in range(agent_isolated.action_dim):
                            dx, dy, dz = agent_isolated.decode_action(a)
                            cand_next = (state[0] + dx, state[1] + dy, state[2] + dz)
                            if is_valid_state(cand_next):
                                legal_actions.append(a)
                        if not legal_actions:
                            break
                        if random.random() < agent_isolated.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            normalized_state = normalize_state(state)
                            state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent_isolated.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        dx, dy, dz = agent_isolated.decode_action(action)
                        next_state = clip_state((state[0] + dx, state[1] + dy, state[2] + dz))
                        triggered = execute_Tr(*next_state)
                        reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                        if target_path.issubset(triggered):
                            reward += 2.0
                        done = (step == N_STEPS - 1)
                        td_error = agent_isolated.store_transition(state, action, reward, next_state, done)

                        priority_stats = agent_isolated.replay_buffer.get_priority_statistics()
                        current_priority = priority_stats['mean_priority'] if priority_stats else 0
                        prioritized_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                current_priority, None)
                        episode_similarities.append(jaccard_similarity(triggered, target_path))
                        episode_td_errors.append(td_error)
                        if prev_reward is not None:
                            prioritized_metrics.record_action_improvement(reward, prev_reward)
                        prev_state = state
                        prev_triggered = triggered
                        prev_reward = reward
                        state = next_state
                        episode_reward += reward

                for _ in range(REPLAY_TIMES):
                    if len(agent_isolated.replay_buffer) >= batch_size:
                        agent_isolated.train(batch_size)
                        global_replay_count += 1
                        if global_replay_count % 2 == 0:
                            agent_isolated.update_target_model()

            if stage1_samples:
                for sample_idx in range(N_SAMPLES_STAGE1):
                    if sample_idx >= len(stage1_samples):
                        break
                    stage1_state_tuple, _, _, _ = stage1_samples[sample_idx]
                    state = stage1_state_tuple
                    prev_state = None
                    prev_triggered = None
                    prev_reward = None

                    for step in range(N_STEPS):
                        legal_actions = []
                        for a in range(agent_isolated.action_dim):
                            dx, dy, dz = agent_isolated.decode_action(a)
                            cand_next = (state[0] + dx, state[1] + dy, state[2] + dz)
                            if is_valid_state(cand_next):
                                legal_actions.append(a)
                        if not legal_actions:
                            break
                        if random.random() < agent_isolated.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            normalized_state = normalize_state(state)
                            state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent_isolated.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        dx, dy, dz = agent_isolated.decode_action(action)
                        next_state = clip_state((state[0] + dx, state[1] + dy, state[2] + dz))
                        triggered = execute_Tr(*next_state)
                        reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                        reward *= 0.8
                        done = (step == N_STEPS - 1)
                        td_error = agent_isolated.store_transition(state, action, reward, next_state, done)

                        priority_stats = agent_isolated.replay_buffer.get_priority_statistics()
                        current_priority = priority_stats['mean_priority'] if priority_stats else 0
                        prioritized_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                current_priority, None)
                        episode_similarities.append(jaccard_similarity(triggered, target_path))
                        episode_td_errors.append(td_error)
                        if prev_reward is not None:
                            prioritized_metrics.record_action_improvement(reward, prev_reward)
                        prev_state = state
                        prev_triggered = triggered
                        prev_reward = reward
                        state = next_state
                        episode_reward += reward

                for _ in range(REPLAY_TIMES):
                    if len(agent_isolated.replay_buffer) >= batch_size:
                        agent_isolated.train(batch_size)
                        global_replay_count += 1
                        if global_replay_count % 2 == 0:
                            agent_isolated.update_target_model()

            trained_paths.add(path_idx)

        avg_similarity = np.mean(episode_similarities) if episode_similarities else 0
        avg_td_error = np.mean(episode_td_errors) if episode_td_errors else 0
        priority_stats = agent_isolated.replay_buffer.get_priority_statistics()
        prioritized_metrics.record_episode_metrics(episode, episode_reward, avg_similarity, avg_td_error,
                                                   agent_isolated.epsilon, "isolated", priority_stats)
        if len(trained_paths) == len(isolated_group):
            break
    return agent_isolated


# === Excel报告生成 ===
def append_metrics_to_combined_excel(metrics_collector, agent_similar, agent_isolated, similar_group, isolated_group,
                                     targetPaths, filepath, run_number):
    training_time = metrics_collector.end_time - metrics_collector.start_time if metrics_collector.end_time else 0
    avg_memory = metrics_collector.total_memory_usage / metrics_collector.memory_check_count if metrics_collector.memory_check_count > 0 else 0
    avg_similarity = np.mean(metrics_collector.final_output_similarities) if metrics_collector.final_output_similarities else 0
    avg_td_error = np.mean(metrics_collector.td_errors) if metrics_collector.td_errors else 0
    action_improve_rate = np.mean(metrics_collector.action_improvements) if metrics_collector.action_improvements else 0
    avg_priority = np.mean(metrics_collector.priority_statistics) if metrics_collector.priority_statistics else 0

    performance_row = {
        'Run': run_number,
        'Average Similarity': f"{avg_similarity:.4f}",
        'TD Error': f"{avg_td_error:.4f}",
        'Action Improve Rate': f"{action_improve_rate:.4f}",
        'Memory(MB)': f"{avg_memory:.2f}",
        'Avg Priority': f"{avg_priority:.4f}",
        'Training Time(s)': f"{training_time:.2f}"
    }

    sample_rows = []
    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_similar.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, reward, sim, triggered in high_reward_samples:
            sample_rows.append({
                'Run': run_number, 'Group Type': 'Similar', 'Path ID': path_idx + 1,
                'X': state_tuple[0], 'Y': state_tuple[1], 'Z': state_tuple[2],
                'Similarity': f"{sim:.4f}", 'Reward': f"{reward:.2f}",
                'Triggered Count': len(triggered), 'Target Count': len(target_path),
                'Triggered Rules': str(sorted(triggered)), 'Target Rules': str(sorted(target_path))
            })
    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, reward, sim, triggered in high_reward_samples:
            sample_rows.append({
                'Run': run_number, 'Group Type': 'Isolated', 'Path ID': path_idx + 1,
                'X': state_tuple[0], 'Y': state_tuple[1], 'Z': state_tuple[2],
                'Similarity': f"{sim:.4f}", 'Reward': f"{reward:.2f}",
                'Triggered Count': len(triggered), 'Target Count': len(target_path),
                'Triggered Rules': str(sorted(triggered)), 'Target Rules': str(sorted(target_path))
            })

    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    if os.path.exists(filepath):
        try:
            df_performance = pd.read_excel(filepath, sheet_name='Performance')
            df_performance = pd.concat([df_performance, pd.DataFrame([performance_row])], ignore_index=True)
        except:
            df_performance = pd.DataFrame([performance_row])
        try:
            df_samples = pd.read_excel(filepath, sheet_name='Final Samples')
            df_samples = pd.concat([df_samples, pd.DataFrame(sample_rows)], ignore_index=True)
        except:
            df_samples = pd.DataFrame(sample_rows)
    else:
        df_performance = pd.DataFrame([performance_row])
        df_samples = pd.DataFrame(sample_rows)

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df_performance.to_excel(writer, sheet_name='Performance', index=False)
        df_samples.to_excel(writer, sheet_name='Final Samples', index=False)
        workbook = writer.book
        ws_performance = writer.sheets['Performance']
        ws_performance.column_dimensions['A'].width = 15
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws_performance.column_dimensions[col].width = 20
        header_font = Font(bold=True, size=11)
        for cell in ws_performance[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in ws_performance.iter_rows(min_row=2, max_row=ws_performance.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        ws_samples = writer.sheets['Final Samples']
        column_widths = {'A': 12, 'B': 15, 'C': 12, 'D': 10, 'E': 10, 'F': 10,
                         'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 40, 'L': 40}
        for col, width in column_widths.items():
            ws_samples.column_dimensions[col].width = width
        for cell in ws_samples[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in ws_samples.iter_rows(min_row=2, max_row=ws_samples.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    print(f"第 {run_number} 次运行结果已保存到: {filepath}")


# === 单次实验运行 ===
def run_single_experiment(run_number, results_save_dir):
    print(f"\n{'=' * 80}")
    print(f"开始第 {run_number} 次运行（优先级DQN）")
    print(f"{'=' * 80}\n")

    prioritized_metrics.reset()
    prioritized_metrics.start_training()

    model_path_similar = os.path.join(results_save_dir, f"run{run_number}_similar.pth")
    model_path_isolated = os.path.join(results_save_dir, f"run{run_number}_isolated.pth")
    path_documents = r"D:\Experiment\CNN\DQNNEW\path_samples_prioritized"

    similar_group, isolated_group = group_paths_by_similarity(targetPaths)

    # 检查样本文件
    def check_sample_files():
        for path_idx in similar_group:
            file_path = os.path.join(path_documents, f"prioritized_path{path_idx + 1}.txt")
            if not os.path.exists(file_path):
                return False
        for path_idx in isolated_group:
            file_path = os.path.join(path_documents, f"prioritized_path{path_idx + 1}_isolated.txt")
            if not os.path.exists(file_path):
                return False
        return True

    print("检查样本文件...")
    if not check_sample_files():
        print("样本文件缺失，正在生成...")
        os.makedirs(path_documents, exist_ok=True)
        print("[1/2] 生成相似路径组样本...")
        generate_samples_for_similar_paths(similar_group, num_total=2000, top_k=200)
        print("相似路径组样本生成完成！")

        temp_replay_buffer = PrioritizedExperienceReplay(capacity=10000)
        temp_agent = PrioritizedDQNAgent(3, 30, temp_replay_buffer)
        print("[2/2] 生成隔离路径组样本...")
        generate_samples_for_isolated_paths_prioritized(temp_agent, isolated_group, num_total=2000, top_k=200)
        print("隔离路径组样本生成完成！")
    else:
        print("所有样本文件已存在。")

    # 阶段1: 相似路径训练
    replay_buffer = PrioritizedExperienceReplay(capacity=10000)
    state_dim = 3
    action_dim = 30
    agent = PrioritizedDQNAgent(state_dim, action_dim, replay_buffer)
    agent = prioritized_generate_and_train_for_similar_paths(agent, similar_group, path_documents, episodes=500,
                                                             batch_size=32, is_isolated=False)
    os.makedirs(os.path.dirname(model_path_similar), exist_ok=True)
    torch.save({
        'model_state_dict': agent.model.state_dict(),
        'optimizer_state_dict': agent.optimizer.state_dict(),
        'epsilon': agent.epsilon
    }, model_path_similar)

    # 阶段2: 隔离路径训练
    isolated_replay_buffer = PrioritizedExperienceReplay(capacity=15000)
    agent_isolated = PrioritizedDQNAgent(state_dim, action_dim, isolated_replay_buffer)
    try:
        checkpoint = torch.load(model_path_similar)
        agent_isolated.model.load_state_dict(checkpoint['model_state_dict'])
        agent_isolated.target_model.load_state_dict(checkpoint['model_state_dict'])
        agent_isolated.epsilon = checkpoint.get('epsilon', 0.5)
        print("成功加载相似路径模型用于隔离路径训练")
    except Exception as e:
        print(f"无法加载相似路径模型: {e}")

    agent_isolated = prioritized_generate_and_train_for_isolated_paths(
        agent_similar=agent, agent_isolated=agent_isolated,
        similar_group=similar_group, isolated_group=isolated_group,
        path_documents=path_documents, episodes=500, batch_size=32, is_isolated=True
    )
    os.makedirs(os.path.dirname(model_path_isolated), exist_ok=True)
    torch.save({
        'model_state_dict': agent_isolated.model.state_dict(),
        'optimizer_state_dict': agent_isolated.optimizer.state_dict(),
        'epsilon': agent_isolated.epsilon
    }, model_path_isolated)

    prioritized_metrics.end_training()

    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, _, sim, triggered in high_reward_samples:
            prioritized_metrics.record_final_output_sample(triggered, target_path)
    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, _, sim, triggered in high_reward_samples:
            prioritized_metrics.record_final_output_sample(triggered, target_path)

    combined_excel_path = os.path.join(results_save_dir, "Prioritized_综合实验结果.xlsx")
    append_metrics_to_combined_excel(
        metrics_collector=prioritized_metrics,
        agent_similar=agent, agent_isolated=agent_isolated,
        similar_group=similar_group, isolated_group=isolated_group,
        targetPaths=targetPaths, filepath=combined_excel_path, run_number=run_number
    )

    avg_similarity = np.mean(prioritized_metrics.final_output_similarities)
    print(f"\n第 {run_number} 次运行完成:")
    print(f"  平均相似度: {avg_similarity:.4f}")
    print(f"  总步数: {prioritized_metrics.step_count}")


if __name__ == "__main__":
    results_save_dir = r"D:\Experiment\CNN\DQNNEW\results\prioritized_results"
    os.makedirs(results_save_dir, exist_ok=True)

    NUM_RUNS = 20
    print("=" * 80)
    print(f"开始 {NUM_RUNS} 次优先级DQN实验")
    print(f"参数范围: X[{X_MIN}, {X_MAX}], Y[{Y_MIN}, {Y_MAX}], Z[{Z_MIN}, {Z_MAX}]")
    print(f"规则总数: 71条")
    print("=" * 80)

    for run in range(1, NUM_RUNS + 1):
        try:
            run_single_experiment(run, results_save_dir)
        except Exception as e:
            print(f"\n第 {run} 次运行出错: {str(e)}")
            import traceback
            traceback.print_exc()
            continue

    print("\n" + "=" * 80)
    print(f" {NUM_RUNS} 次运行全部完成")
    print(f"结果保存目录: {results_save_dir}")
    print("=" * 80)