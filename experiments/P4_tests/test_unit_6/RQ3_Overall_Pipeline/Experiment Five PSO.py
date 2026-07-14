import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """安全除法"""
    if denominator == 0:
        return default
    return numerator / denominator


def execute_Tr(x, y, z):
    triggered = set()

    # --- 分支 1-15 (原 process_b * process_c / (process_a + 1) > 110 的变异) ---
    if ((y * z) / (x + 1) > 110) != ((x * z) / (x + 1) > 110): triggered.add(1)
    if ((y * z) / (x + 1) > 110) != ((z * z) / (x + 1) > 110): triggered.add(2)
    if ((y * z) / (x + 1) > 110) != ((y * y) / (x + 1) > 110): triggered.add(3)
    if ((y * z) / (x + 1) > 110) != ((y * x) / (x + 1) > 110): triggered.add(4)
    if ((y * z) / (x + 1) > 110) != ((y * 12) / (x + 1) > 110): triggered.add(5)
    if ((y * z) / (x + 1) > 110) != ((y * z) / (y + 1) > 110): triggered.add(6)
    if ((y * z) / (x + 1) > 110) != ((y * z) / (z + 1) > 110): triggered.add(7)
    if ((y * z) / (x + 1) > 110) != ((y * z) / (x + 5) > 110): triggered.add(8)
    if ((y * z) / (x + 1) > 110) != ((y * z) / (x + 1) > 130): triggered.add(9)
    if ((y * z) / (x + 1) > 110) != ((50 * z) / (x + 1) > 110): triggered.add(10)
    if ((y * z) / (x + 1) > 110) != ((y * z) / (x - 1) > 110): triggered.add(11)
    if ((y * z) / (x + 1) > 110) != ((y * z * 3) / (x + 1) > 110): triggered.add(12)
    if ((y * z) / (x + 1) > 110) != ((y * z) / (x * 0.5 + 1) > 110): triggered.add(13)
    if ((y * z) / (x + 1) > 110) != ((y * z) / (x * 2 + 1) > 110): triggered.add(14)
    if ((y * z) / (x + 1) > 110) != ((y * z) / (60 + 1) > 110): triggered.add(15)

    # --- 分支 16-26 (原 (process_c - process_a) < 0.28 * process_b 的变异) ---
    if ((z - x) < 0.28 * y) != ((y - x) < 0.28 * y): triggered.add(16)
    if ((z - x) < 0.28 * y) != ((z * 2 - x) < 0.28 * y): triggered.add(17)
    if ((z - x) < 0.28 * y) != ((z * 1.5 - x) < 0.28 * y): triggered.add(18)
    if ((z - x) < 0.28 * y) != ((z - y) < 0.28 * y): triggered.add(19)
    if ((z - x) < 0.28 * y) != ((z - x) < 0.28 * x): triggered.add(20)
    if ((z - x) < 0.28 * y) != ((z - x) < 0.28 * z): triggered.add(21)
    if ((z - x) < 0.28 * y) != ((z - x) < 0.48 * y): triggered.add(22)
    if ((z - x) < 0.28 * y) != ((z - x) < 0.18 * y): triggered.add(23)
    if ((z - x) < 0.28 * y) != ((z + x) < 0.28 * y): triggered.add(24)
    if ((z - x) < 0.28 * y) != ((z - x * 1.2) < 0.28 * y): triggered.add(25)
    if ((z - x) < 0.28 * y) != ((z - x * 2) < 0.28 * y): triggered.add(26)

    # --- 分支 27-41 (原 (process_a^3 + process_b^3) < process_c^2 的变异) ---
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2.5 + y ** 3) < z ** 2): triggered.add(27)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 2.5) < z ** 2): triggered.add(28)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 - y ** 3) < z ** 2): triggered.add(29)
    if ((x ** 3 + y ** 3) < z ** 2) != ((y ** 3 + y ** 3) < z ** 2): triggered.add(30)
    if ((x ** 3 + y ** 3) < z ** 2) != ((z ** 3 + y ** 3) < z ** 2): triggered.add(31)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + x ** 3) < z ** 2): triggered.add(32)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + z ** 3) < z ** 2): triggered.add(33)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 2.5) < z ** 2): triggered.add(34)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < x ** 2): triggered.add(35)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < y ** 2): triggered.add(36)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 2.3): triggered.add(37)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 1.5): triggered.add(38)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * 3) < z ** 2): triggered.add(39)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x * 3 + y ** 3) < z ** 2): triggered.add(40)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z * 2): triggered.add(41)

    # --- 分支 42-54 (原 abs(process_c-(process_a+process_b))<2.5 and abs(process_b-process_a*1.25)<1.5 的变异) ---
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z * 2 - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5): triggered.add(42)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + x)) < 2.5 and abs(y - x * 1.25) < 1.5): triggered.add(43)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x - y)) < 2.5 and abs(y - x * 1.25) < 1.5): triggered.add(44)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (y + y)) < 2.5 and abs(y - x * 1.25) < 1.5): triggered.add(45)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x * 0.9 + y)) < 2.5 and abs(y - x * 1.25) < 1.5): triggered.add(46)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + y * 0.8)) < 2.5 and abs(y - x * 1.25) < 1.5): triggered.add(47)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + y)) < 7 and abs(y - x * 1.25) < 1.5): triggered.add(48)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + y)) < 2.5 and abs(z - x * 1.25) < 1.5): triggered.add(49)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + y)) < 2.5 and abs(y * 1.5 - x * 1.25) < 1.5): triggered.add(50)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + y)) < 2.5 and abs(y - x * 1.75) < 1.5): triggered.add(51)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + y)) < 2.5 and abs(y - z * 1.25) < 1.5): triggered.add(52)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 8): triggered.add(53)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + y)) < 2.5 and abs(y - x ** 1.25) < 1.5): triggered.add(54)

    # --- 分支 55-67 (原 x/(y+0.01)>3.5 and y/(z+0.01)<0.3 的变异) ---
    # 注意：原代码中分支56和66都赋值67，已修正
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x * 1.3 / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3): triggered.add(55)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y * 0.6 + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3): triggered.add(56)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y + 0.01)) > 8 and (y / (z + 0.01)) < 0.3): triggered.add(57)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y + 0.01)) > 3.5 and (z / (z + 0.01)) < 0.3): triggered.add(58)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y + 0.01)) > 3.5 and (x / (z + 0.01)) < 0.3): triggered.add(59)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((y / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3): triggered.add(60)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((z / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3): triggered.add(61)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x % (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3): triggered.add(62)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y + 0.01)) > 3.5 and (y % (z + 0.01)) < 0.3): triggered.add(63)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y + 0.01)) > 3.5 or (y / (z + 0.01)) < 0.3): triggered.add(64)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.2): triggered.add(65)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 1.3): triggered.add(66)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y + 0.01)) > 3.5 and (y / (z * 0.7 + 0.01)) < 0.3): triggered.add(67)

    # --- 分支 68-81 (原 abs(x-y)>14 and abs(y-z)>16 and abs(x-z)<7 的变异) ---
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x * 1.2 - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7): triggered.add(68)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x * 1.6 - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7): triggered.add(69)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y * 1.2) > 14 and abs(y - z) > 16 and abs(x - z) < 7): triggered.add(70)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - z) > 14 and abs(y - z) > 16 and abs(x - z) < 7): triggered.add(71)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(x - z) > 16 and abs(x - z) < 7): triggered.add(72)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(y - x) > 16 and abs(x - z) < 7): triggered.add(73)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(y - z) > 16 and abs(y - z) < 7): triggered.add(74)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - y) < 7): triggered.add(75)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 12): triggered.add(76)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(y - z) > 16 and abs(x * 2 - z) < 7): triggered.add(77)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(y * 3 - z) > 16 and abs(x - z) < 7): triggered.add(78)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(y - z * 1.6) > 16 and abs(x - z) < 7): triggered.add(79)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(y - z * 0.7) > 16 and abs(x - z) < 7): triggered.add(80)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 20 and abs(y - z) > 16 and abs(x - z) < 7): triggered.add(81)

    # --- 分支 82-93 (原 (x>85 or x<8) and (y>80 or y<5) and (z>75 or z<4) 的变异) ---
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x * x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(82)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x * y > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(83)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x * z > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(84)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x > 15 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(85)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x > 85 or x < 8) and (y * x > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(86)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x > 85 or x < 8) and (y * y > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(87)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x > 85 or x < 8) and (y * z > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(88)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x > 85 or x < 8) and (y > 10 or y < 5) and (z > 75 or z < 4)): triggered.add(89)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x * 50 > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(90)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x * 80 > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(91)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x > 15 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(92)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x > 85 or x < 8) and (y > 80 or y < 5) and (z * z > 75 or z < 4)): triggered.add(93)

    # --- 分支 94-107 (原 (x+y)^1.2<z^1.8 and (x+y+z)/3>45 的变异) ---
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((y + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45): triggered.add(94)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((z + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45): triggered.add(95)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + x) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45): triggered.add(96)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + z) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45): triggered.add(97)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.8 < z ** 1.8 and (x + y + z) / 3 > 45): triggered.add(98)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.2 < x ** 1.8 and (x + y + z) / 3 > 45): triggered.add(99)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.2 < y ** 1.8 and (x + y + z) / 3 > 45): triggered.add(100)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 0.8 < z ** 1.8 and (x + y + z) / 3 > 45): triggered.add(101)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.2 < z ** 1.8 and (x + x + z) / 3 > 45): triggered.add(102)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.2 < z ** 1.8 and (x + z + z) / 3 > 45): triggered.add(103)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.2 < z ** 1.8 and (x + y + x) / 3 > 45): triggered.add(104)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.2 < z ** 1.8 and (x + y + y) / 3 > 45): triggered.add(105)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.2 < z ** 1.8 and (x + 50 + z) / 3 > 45): triggered.add(106)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.2 < z ** 1.8 and (60 + y + z) / 3 > 45): triggered.add(107)

    return triggered


def execute_validation_rules(dx: int, dy: int, dz: int) -> Set[int]:
    """执行验证规则，返回触发的分支集合（调用105分支版本）"""
    return execute_Tr(dx, dy, dz)


def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """计算适应度"""
    generated_path = execute_validation_rules(int(particle[0]), int(particle[1]), int(particle[2]))

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """基本PSO优化器"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # 范围: x(2,100), y(20,150), z(30,200)
        self.bounds = bounds if bounds else [(2, 100), (20, 150), (30, 200)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """初始化粒子"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """更新速度和位置"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """对目标路径进行PSO优化"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(
                            int(particles[i][0]), int(particles[i][1]), int(particles[i][2])
                        ),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(
                int(gbest_particle[0]), int(gbest_particle[1]), int(gbest_particle[2])
            ),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """对所有路径运行PSO"""

    print(f"\n{'=' * 70}")
    print(f"Baseline PSO - 路径优化")
    print(f"{'=' * 70}")
    print(f"粒子数: {n_particles}, 最大迭代次数: {max_iterations}")
    print(f"路径数量: {len(target_paths)}")
    print(f"范围: x(2,100), y(1,105), z(1,110)")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"路径 {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "完美求解" if result['success'] else f"(适应度: {result['best_fitness']:.3f})"
        print(f"{status} | 耗时: {result['time']:.2f}s | 迭代次数: {result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f"求解成功: {success_count}/{len(target_paths)} ({success_rate:.1f}%) | 总耗时: {total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """多次运行实验"""

    print(f"\n{'=' * 70}")
    print(f"Baseline PSO - {num_runs}次运行")
    print(f"{'=' * 70}")
    print(f"粒子数: {n_particles}, 最大迭代次数: {max_iterations}, 路径数: {len(target_paths)}")
    print(f"范围: x(2,100), y(1,105), z(1,110)")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"--- 第 {run_idx}/{num_runs} 次运行 ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f"求解成功: {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs}次运行完成 | 总耗时: {total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """导出Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # Sheet 1: 运行汇总
    ws1 = wb.active
    ws1.title = "运行汇总"
    ws1.sheet_view.showGridLines = False

    headers = ["运行", "成功率", "求解/总数", "平均适应度", "平均迭代次数", "耗时(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f"运行 {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # Sheet 2: 路径统计
    ws2 = wb.create_sheet(title="路径统计")
    ws2.sheet_view.showGridLines = False

    headers2 = ["路径ID", "求解/总数", "成功率", "平均适应度", "平均迭代次数", "最小迭代次数", "最大迭代次数"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"路径 {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # Sheet 3: 详细结果
    ws3 = wb.create_sheet(title="详细结果")
    ws3.sheet_view.showGridLines = False

    headers3 = ["路径ID", "运行", "位置(x,y,z)", "适应度", "迭代次数", "触发路径"]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"Run {run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # Sheet 4: 目标路径
    ws4 = wb.create_sheet(title="目标路径")
    ws4.sheet_view.showGridLines = False

    headers4 = ["路径ID", "目标路径", "分支数"]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"路径 {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f"Excel已保存: {filename}")
    print(f"{'=' * 70}")
    print(f"包含工作表:")
    print(f"  1. 运行汇总      - {len(all_results)}次运行汇总")
    print(f"  2. 路径统计      - 每条路径的统计")
    print(f"  3. 详细结果      - 每次运行每条路径的详细结果")
    print(f"  4. 目标路径      - 目标路径定义")
    print(f"{'=' * 70}\n")

    return filename


def main():
    # 目标路径定义（105分支版本）
    target_paths = [
    {1, 2, 5, 6, 8, 9, 10, 14, 16, 17, 18, 24, 29, 69, 71, 72, 74, 75, 77, 79, 82, 83, 84, 85, 90, 91, 92, 98},
    {1, 2, 5, 6, 8, 9, 10, 14, 16, 17, 18, 24, 29, 71, 72, 74, 75, 77, 79, 82, 83, 84, 85, 90, 91, 92, 98},
    {1, 2, 4, 5, 6, 7, 8, 9, 10, 14, 17, 18, 24, 29, 71, 72, 74, 75, 77, 82, 83, 84, 85, 90, 91, 92, 98},
    {3, 4, 7, 12, 13, 15, 17, 18, 24, 29, 68, 71, 72, 74, 75, 77, 79, 82, 83, 84, 85, 90, 91, 92, 98},
    {1, 2, 4, 5, 6, 7, 10, 14, 15, 16, 17, 18, 20, 21, 24, 29, 77, 94, 98, 99, 102, 103, 104, 106},
    {1, 3, 4, 5, 7, 15, 30, 31, 33, 35, 36, 38, 41, 61, 62, 64, 86, 87, 88, 89, 99, 102, 104, 105},
    {4, 5, 7, 16, 26, 29, 42, 43, 44, 45, 46, 47, 49, 50, 51, 52, 54, 82, 83, 84, 85, 90, 91, 92},
    {1, 3, 4, 5, 6, 7, 8, 9, 10, 14, 16, 19, 21, 22, 25, 26, 29, 82, 83, 84, 85, 90, 91, 92, 98},
    {1, 4, 6, 7, 15, 16, 17, 18, 20, 21, 23, 24, 29, 32, 39, 93, 94, 98, 99, 102, 103, 104, 106},
    {1, 4, 5, 6, 7, 15, 28, 29, 32, 34, 37, 39, 61, 62, 64, 86, 87, 88, 89, 99, 102, 104, 105},
    {16, 20, 21, 22, 25, 26, 57, 58, 59, 60, 62, 63, 65, 67, 76, 86, 87, 88, 89, 98, 100, 105},
    {2, 16, 20, 21, 25, 26, 40, 49, 57, 58, 59, 60, 63, 65, 86, 87, 88, 89, 98, 100, 105},
    {3, 4, 5, 7, 8, 9, 14, 15, 16, 27, 28, 29, 30, 34, 37, 39, 40, 61, 62, 64, 104, 105},
    {1, 3, 4, 5, 6, 7, 8, 9, 14, 15, 28, 29, 32, 34, 37, 39, 61, 62, 64, 103, 106, 107},
    {1, 2, 5, 6, 9, 10, 14, 15, 16, 24, 29, 71, 72, 74, 75, 77, 96, 97, 99, 100, 101},
    {12, 13, 15, 17, 18, 19, 24, 70, 71, 72, 74, 75, 77, 80, 81, 86, 87, 88, 89, 98},
    {3, 12, 13, 15, 17, 18, 24, 29, 73, 78, 79, 80, 82, 83, 84, 85, 90, 91, 92, 98},
    {1, 6, 15, 16, 17, 18, 20, 21, 29, 32, 39, 93, 94, 95, 98, 99, 102, 103, 106},
    {3, 4, 7, 11, 12, 13, 15, 17, 18, 24, 29, 68, 71, 72, 73, 74, 75, 77, 81, 98},
    {2, 16, 20, 21, 22, 25, 26, 55, 56, 61, 62, 64, 76, 86, 87, 88, 89, 98},
    {4, 5, 7, 10, 14, 16, 26, 29, 46, 48, 82, 83, 84, 85, 90, 91, 92, 98},
    {3, 4, 5, 7, 10, 14, 16, 26, 29, 53, 82, 83, 84, 85, 90, 91, 92},
    {17, 18, 19, 24, 64, 66, 76, 86, 87, 88, 89, 98}  # A18
    ]

    print("=" * 70)
    print("Baseline PSO - 105分支版本")
    print("=" * 70)
    print(f"范围: x(2,100), y(1,105), z(1,110)")
    print(f"路径数: {len(target_paths)}")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("程序完成")


if __name__ == "__main__":
    main()