import os
import random
from collections import deque
import numpy as np
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# --- 全局状态范围配置 ---
STATE_MIN_X, STATE_MAX_X = 2, 100
STATE_MIN_Y, STATE_MAX_Y = 20, 150
STATE_MIN_Z, STATE_MAX_Z = 30, 200

def generate_input():
    return [
        random.randint(STATE_MIN_X, STATE_MAX_X),
        random.randint(STATE_MIN_Y, STATE_MAX_Y),
        random.randint(STATE_MIN_Z, STATE_MAX_Z)
    ]

def execute_Tr(x, y, z):
    triggered = set()

    # --- 分支 1-15 (原 process_b * process_c / (process_a + 1) > 110 的变异) ---
    if ((y * z) / (x + 1) > 110) != ((x * z) / (x + 1) > 110): triggered.add(1)
    if ((y * z) / (x + 1) > 110) != ((z * z) / (x + 1) > 110): triggered.add(2)
    if ((y * z) / (x + 1) > 110) != ((y * y) / (x + 1) > 110): triggered.add(3)
    if ((y * z) / (x + 1) > 110) != ((y * x) / (x + 1) > 110): triggered.add(4)
    if ((y * z) / (x + 1) > 110) != ((y * 12) / (x + 1) > 110): triggered.add(5)
    if ((y * z) / (x + 1) > 110) != ((y * z) / (y + 1) > 110): triggered.add(6)
    if ((y * z) / (x + 1) > 110) != ((y * z) / (z + 1) > 110): triggered.add(7)
    if ((y * z) / (x + 1) > 110) != ((y * z) / (x + 5) > 110): triggered.add(8)
    if ((y * z) / (x + 1) > 110) != ((y * z) / (x + 1) > 130): triggered.add(9)
    if ((y * z) / (x + 1) > 110) != ((50 * z) / (x + 1) > 110): triggered.add(10)
    if ((y * z) / (x + 1) > 110) != ((y * z) / (x - 1) > 110): triggered.add(11)
    if ((y * z) / (x + 1) > 110) != ((y * z * 3) / (x + 1) > 110): triggered.add(12)
    if ((y * z) / (x + 1) > 110) != ((y * z) / (x * 0.5 + 1) > 110): triggered.add(13)
    if ((y * z) / (x + 1) > 110) != ((y * z) / (x * 2 + 1) > 110): triggered.add(14)
    if ((y * z) / (x + 1) > 110) != ((y * z) / (60 + 1) > 110): triggered.add(15)

    # --- 分支 16-26 (原 (process_c - process_a) < 0.28 * process_b 的变异) ---
    if ((z - x) < 0.28 * y) != ((y - x) < 0.28 * y): triggered.add(16)
    if ((z - x) < 0.28 * y) != ((z * 2 - x) < 0.28 * y): triggered.add(17)
    if ((z - x) < 0.28 * y) != ((z * 1.5 - x) < 0.28 * y): triggered.add(18)
    if ((z - x) < 0.28 * y) != ((z - y) < 0.28 * y): triggered.add(19)
    if ((z - x) < 0.28 * y) != ((z - x) < 0.28 * x): triggered.add(20)
    if ((z - x) < 0.28 * y) != ((z - x) < 0.28 * z): triggered.add(21)
    if ((z - x) < 0.28 * y) != ((z - x) < 0.48 * y): triggered.add(22)
    if ((z - x) < 0.28 * y) != ((z - x) < 0.18 * y): triggered.add(23)
    if ((z - x) < 0.28 * y) != ((z + x) < 0.28 * y): triggered.add(24)
    if ((z - x) < 0.28 * y) != ((z - x * 1.2) < 0.28 * y): triggered.add(25)
    if ((z - x) < 0.28 * y) != ((z - x * 2) < 0.28 * y): triggered.add(26)

    # --- 分支 27-41 (原 (process_a^3 + process_b^3) < process_c^2 的变异) ---
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2.5 + y ** 3) < z ** 2): triggered.add(27)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 2.5) < z ** 2): triggered.add(28)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 - y ** 3) < z ** 2): triggered.add(29)
    if ((x ** 3 + y ** 3) < z ** 2) != ((y ** 3 + y ** 3) < z ** 2): triggered.add(30)
    if ((x ** 3 + y ** 3) < z ** 2) != ((z ** 3 + y ** 3) < z ** 2): triggered.add(31)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + x ** 3) < z ** 2): triggered.add(32)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + z ** 3) < z ** 2): triggered.add(33)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 2.5) < z ** 2): triggered.add(34)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < x ** 2): triggered.add(35)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < y ** 2): triggered.add(36)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 2.3): triggered.add(37)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 1.5): triggered.add(38)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * 3) < z ** 2): triggered.add(39)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x * 3 + y ** 3) < z ** 2): triggered.add(40)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z * 2): triggered.add(41)

    # --- 分支 42-54 (原 abs(process_c-(process_a+process_b))<2.5 and abs(process_b-process_a*1.25)<1.5 的变异) ---
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z * 2 - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5): triggered.add(42)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + x)) < 2.5 and abs(y - x * 1.25) < 1.5): triggered.add(43)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x - y)) < 2.5 and abs(y - x * 1.25) < 1.5): triggered.add(44)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (y + y)) < 2.5 and abs(y - x * 1.25) < 1.5): triggered.add(45)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x * 0.9 + y)) < 2.5 and abs(y - x * 1.25) < 1.5): triggered.add(46)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + y * 0.8)) < 2.5 and abs(y - x * 1.25) < 1.5): triggered.add(47)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + y)) < 7 and abs(y - x * 1.25) < 1.5): triggered.add(48)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + y)) < 2.5 and abs(z - x * 1.25) < 1.5): triggered.add(49)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + y)) < 2.5 and abs(y * 1.5 - x * 1.25) < 1.5): triggered.add(50)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + y)) < 2.5 and abs(y - x * 1.75) < 1.5): triggered.add(51)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + y)) < 2.5 and abs(y - z * 1.25) < 1.5): triggered.add(52)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 8): triggered.add(53)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + y)) < 2.5 and abs(y - x ** 1.25) < 1.5): triggered.add(54)

    # --- 分支 55-67 (原 x/(y+0.01)>3.5 and y/(z+0.01)<0.3 的变异) ---
    # 注意：原代码中分支56和66都赋值67，已修正
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x * 1.3 / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3): triggered.add(55)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y * 0.6 + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3): triggered.add(56)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y + 0.01)) > 8 and (y / (z + 0.01)) < 0.3): triggered.add(57)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y + 0.01)) > 3.5 and (z / (z + 0.01)) < 0.3): triggered.add(58)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y + 0.01)) > 3.5 and (x / (z + 0.01)) < 0.3): triggered.add(59)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((y / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3): triggered.add(60)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((z / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3): triggered.add(61)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x % (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3): triggered.add(62)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y + 0.01)) > 3.5 and (y % (z + 0.01)) < 0.3): triggered.add(63)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y + 0.01)) > 3.5 or (y / (z + 0.01)) < 0.3): triggered.add(64)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.2): triggered.add(65)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 1.3): triggered.add(66)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y + 0.01)) > 3.5 and (y / (z * 0.7 + 0.01)) < 0.3): triggered.add(67)

    # --- 分支 68-81 (原 abs(x-y)>14 and abs(y-z)>16 and abs(x-z)<7 的变异) ---
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x * 1.2 - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7): triggered.add(68)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x * 1.6 - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7): triggered.add(69)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y * 1.2) > 14 and abs(y - z) > 16 and abs(x - z) < 7): triggered.add(70)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - z) > 14 and abs(y - z) > 16 and abs(x - z) < 7): triggered.add(71)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(x - z) > 16 and abs(x - z) < 7): triggered.add(72)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(y - x) > 16 and abs(x - z) < 7): triggered.add(73)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(y - z) > 16 and abs(y - z) < 7): triggered.add(74)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - y) < 7): triggered.add(75)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 12): triggered.add(76)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(y - z) > 16 and abs(x * 2 - z) < 7): triggered.add(77)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(y * 3 - z) > 16 and abs(x - z) < 7): triggered.add(78)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(y - z * 1.6) > 16 and abs(x - z) < 7): triggered.add(79)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(y - z * 0.7) > 16 and abs(x - z) < 7): triggered.add(80)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 20 and abs(y - z) > 16 and abs(x - z) < 7): triggered.add(81)

    # --- 分支 82-93 (原 (x>85 or x<8) and (y>80 or y<5) and (z>75 or z<4) 的变异) ---
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x * x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(82)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x * y > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(83)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x * z > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(84)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x > 15 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(85)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x > 85 or x < 8) and (y * x > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(86)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x > 85 or x < 8) and (y * y > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(87)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x > 85 or x < 8) and (y * z > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(88)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x > 85 or x < 8) and (y > 10 or y < 5) and (z > 75 or z < 4)): triggered.add(89)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x * 50 > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(90)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x * 80 > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(91)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x > 15 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(92)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x > 85 or x < 8) and (y > 80 or y < 5) and (z * z > 75 or z < 4)): triggered.add(93)

    # --- 分支 94-107 (原 (x+y)^1.2<z^1.8 and (x+y+z)/3>45 的变异) ---
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((y + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45): triggered.add(94)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((z + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45): triggered.add(95)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + x) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45): triggered.add(96)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + z) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45): triggered.add(97)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.8 < z ** 1.8 and (x + y + z) / 3 > 45): triggered.add(98)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.2 < x ** 1.8 and (x + y + z) / 3 > 45): triggered.add(99)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.2 < y ** 1.8 and (x + y + z) / 3 > 45): triggered.add(100)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 0.8 < z ** 1.8 and (x + y + z) / 3 > 45): triggered.add(101)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.2 < z ** 1.8 and (x + x + z) / 3 > 45): triggered.add(102)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.2 < z ** 1.8 and (x + z + z) / 3 > 45): triggered.add(103)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.2 < z ** 1.8 and (x + y + x) / 3 > 45): triggered.add(104)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.2 < z ** 1.8 and (x + y + y) / 3 > 45): triggered.add(105)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.2 < z ** 1.8 and (x + 50 + z) / 3 > 45): triggered.add(106)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.2 < z ** 1.8 and (60 + y + z) / 3 > 45): triggered.add(107)

    return triggered


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0


targetPaths = [
    {1, 2, 5, 6, 8, 9, 10, 14, 16, 17, 18, 24, 29, 69, 71, 72, 74, 75, 77, 79, 82, 83, 84, 85, 90, 91, 92, 98},
    {1, 2, 5, 6, 8, 9, 10, 14, 16, 17, 18, 24, 29, 71, 72, 74, 75, 77, 79, 82, 83, 84, 85, 90, 91, 92, 98},
    {1, 2, 4, 5, 6, 7, 8, 9, 10, 14, 17, 18, 24, 29, 71, 72, 74, 75, 77, 82, 83, 84, 85, 90, 91, 92, 98},
    {3, 4, 7, 12, 13, 15, 17, 18, 24, 29, 68, 71, 72, 74, 75, 77, 79, 82, 83, 84, 85, 90, 91, 92, 98},
    {1, 2, 4, 5, 6, 7, 10, 14, 15, 16, 17, 18, 20, 21, 24, 29, 77, 94, 98, 99, 102, 103, 104, 106},
    {1, 3, 4, 5, 7, 15, 30, 31, 33, 35, 36, 38, 41, 61, 62, 64, 86, 87, 88, 89, 99, 102, 104, 105},
    {4, 5, 7, 16, 26, 29, 42, 43, 44, 45, 46, 47, 49, 50, 51, 52, 54, 82, 83, 84, 85, 90, 91, 92},
    {1, 3, 4, 5, 6, 7, 8, 9, 10, 14, 16, 19, 21, 22, 25, 26, 29, 82, 83, 84, 85, 90, 91, 92, 98},
    {1, 4, 6, 7, 15, 16, 17, 18, 20, 21, 23, 24, 29, 32, 39, 93, 94, 98, 99, 102, 103, 104, 106},
    {1, 4, 5, 6, 7, 15, 28, 29, 32, 34, 37, 39, 61, 62, 64, 86, 87, 88, 89, 99, 102, 104, 105},
    {16, 20, 21, 22, 25, 26, 57, 58, 59, 60, 62, 63, 65, 67, 76, 86, 87, 88, 89, 98, 100, 105},
    {2, 16, 20, 21, 25, 26, 40, 49, 57, 58, 59, 60, 63, 65, 86, 87, 88, 89, 98, 100, 105},
    {3, 4, 5, 7, 8, 9, 14, 15, 16, 27, 28, 29, 30, 34, 37, 39, 40, 61, 62, 64, 104, 105},
    {1, 3, 4, 5, 6, 7, 8, 9, 14, 15, 28, 29, 32, 34, 37, 39, 61, 62, 64, 103, 106, 107},
    {1, 2, 5, 6, 9, 10, 14, 15, 16, 24, 29, 71, 72, 74, 75, 77, 96, 97, 99, 100, 101},
    {12, 13, 15, 17, 18, 19, 24, 70, 71, 72, 74, 75, 77, 80, 81, 86, 87, 88, 89, 98},
    {3, 12, 13, 15, 17, 18, 24, 29, 73, 78, 79, 80, 82, 83, 84, 85, 90, 91, 92, 98},
    {1, 6, 15, 16, 17, 18, 20, 21, 29, 32, 39, 93, 94, 95, 98, 99, 102, 103, 106},
    {3, 4, 7, 11, 12, 13, 15, 17, 18, 24, 29, 68, 71, 72, 73, 74, 75, 77, 81, 98},
    {2, 16, 20, 21, 22, 25, 26, 55, 56, 61, 62, 64, 76, 86, 87, 88, 89, 98},
    {4, 5, 7, 10, 14, 16, 26, 29, 46, 48, 82, 83, 84, 85, 90, 91, 92, 98},
    {3, 4, 5, 7, 10, 14, 16, 26, 29, 53, 82, 83, 84, 85, 90, 91, 92},
    {17, 18, 19, 24, 64, 66, 76, 86, 87, 88, 89, 98}
]


class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 6000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(targetPaths)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }


def compute_robustness(state, path):
    base = execute_Tr(*state)
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                        np.clip(state[0] + dx, STATE_MIN_X, STATE_MAX_X),
                        np.clip(state[1] + dy, STATE_MIN_Y, STATE_MAX_Y),
                        np.clip(state[2] + dz, STATE_MIN_Z, STATE_MAX_Z)
                    ])
                n_trig = execute_Tr(*neighbor)
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = targetPaths[target_path_idx]
    samples = []
    attempts = 0

    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(STATE_MIN_X, STATE_MAX_X),
            random.randint(STATE_MIN_Y, STATE_MAX_Y),
            random.randint(STATE_MIN_Z, STATE_MAX_Z)
        ])
        triggered = execute_Tr(*state)

        if not triggered:
            continue

        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)

        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }

        samples.append(sample_data)

    return samples


def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample[
                'robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)

        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]

    return selected_samples


def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = targetPaths[target_path_idx]
        samples = []
        attempts = 0

        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(STATE_MIN_X, STATE_MAX_X),
                random.randint(STATE_MIN_Y, STATE_MAX_Y),
                random.randint(STATE_MIN_Z, STATE_MAX_Z)
            ])
            triggered = execute_Tr(*state)

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)

            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)

        random.shuffle(samples)
        return samples[:config.top_k_samples]

    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")

        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)


def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}

    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates

    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}

        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config,
                                                         shared_candidates[path_idx])

            strategy_results[path_idx] = samples

        results[strategy_name] = strategy_results

    return results


def analyze_fitness_values(results, config):
    analysis_results = {}

    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []

        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])

        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }

        all_scores = []

        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])

        scores_array = np.array(all_scores)

        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)

        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })

        analysis_results[strategy_name] = analysis

    return analysis_results


def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []

    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)

    df = pd.DataFrame(df_data)
    return df, analysis_results


def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []

    print(f"Starting {num_runs} experiments...")

    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")

        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)

        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)

        print(f"Experiment {run_idx + 1} completed.")

    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df


def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')

    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width

    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []

    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]

        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)

    stats_df = pd.DataFrame(stats_data)

    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Results saved to: {output_path}")


def main():
    results_df = run_multiple_experiments(num_runs=20)

    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")

    save_results_to_excel(results_df, output_path)

    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)

    return results_df, output_path


if __name__ == "__main__":
    results, output_file = main()