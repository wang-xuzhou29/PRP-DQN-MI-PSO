import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === ()===
MIN_X = 2
MAX_X = 100
MIN_Y = 20
MAX_Y = 150
MIN_Z = 30
MAX_Z = 200

# === / ===
def normalize_state(state):
    """将状态归一化到 [0, 1] 区间"""
    weather_norm = (state[0] - MIN_X) / (MAX_X - MIN_X)
    time_norm = (state[1] - MIN_Y) / (MAX_Y - MIN_Y)
    z_norm = (state[2] - MIN_Z) / (MAX_Z - MIN_Z)
    return (weather_norm, time_norm, z_norm)

def denormalize_state(state_norm):
    """将归一化状态还原"""
    weather = int(round(state_norm[0] * (MAX_X - MIN_X) + MIN_X))
    time_period = int(round(state_norm[1] * (MAX_Y - MIN_Y) + MIN_Y))
    z = int(round(state_norm[2] * (MAX_Z - MIN_Z) + MIN_Z))

    # 边界保护
    weather = np.clip(weather, MIN_X, MAX_X)
    time_period = np.clip(time_period, MIN_Y, MAX_Y)
    z = np.clip(z, MIN_Z, MAX_Z)

    return (weather, time_period, z)

def normalize_value(value, min_val, max_val):
    """"""
    return (value - min_val) / (max_val - min_val)

def denormalize_value(value_norm, min_val, max_val):
    """"""
    return int(round(value_norm * (max_val - min_val) + min_val))

# ===  ===
def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

# === reward function ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    """"""
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10

    if target_path.issubset(triggered):
        reward += 1

    if prev_triggered is not None:
        prev_sim = jaccard_similarity(prev_triggered, target_path)
        improvement = sim - prev_sim
        reward += improvement * 5

    return reward

def execute_Tr(x, y, z):
    # 初始化分支覆盖数组
    triggered = set()

    # --- 分支 1-15 (原 process_b * process_c / (process_a + 1) > 110 的变异) ---
    if ((y * z) / (x + 1) > 110) != ((x * z) / (x + 1) > 110): triggered.add(1)
    if ((y * z) / (x + 1) > 110) != ((z * z) / (x + 1) > 110): triggered.add(2)
    if ((y * z) / (x + 1) > 110) != ((y * y) / (x + 1) > 110): triggered.add(3)
    if ((y * z) / (x + 1) > 110) != ((y * x) / (x + 1) > 110): triggered.add(4)
    if ((y * z) / (x + 1) > 110) != ((y * 12) / (x + 1) > 110): triggered.add(5)
    if ((y * z) / (x + 1) > 110) != ((y * z) / (y + 1) > 110): triggered.add(6)
    if ((y * z) / (x + 1) > 110) != ((y * z) / (z + 1) > 110): triggered.add(7)
    if ((y * z) / (x + 1) > 110) != ((y * z) / (x + 5) > 110): triggered.add(8)
    if ((y * z) / (x + 1) > 110) != ((y * z) / (x + 1) > 130): triggered.add(9)
    if ((y * z) / (x + 1) > 110) != ((50 * z) / (x + 1) > 110): triggered.add(10)
    if ((y * z) / (x + 1) > 110) != ((y * z) / (x - 1) > 110): triggered.add(11)
    if ((y * z) / (x + 1) > 110) != ((y * z * 3) / (x + 1) > 110): triggered.add(12)
    if ((y * z) / (x + 1) > 110) != ((y * z) / (x * 0.5 + 1) > 110): triggered.add(13)
    if ((y * z) / (x + 1) > 110) != ((y * z) / (x * 2 + 1) > 110): triggered.add(14)
    if ((y * z) / (x + 1) > 110) != ((y * z) / (60 + 1) > 110): triggered.add(15)

    # --- 分支 16-26 (原 (process_c - process_a) < 0.28 * process_b 的变异) ---
    if ((z - x) < 0.28 * y) != ((y - x) < 0.28 * y): triggered.add(16)
    if ((z - x) < 0.28 * y) != ((z * 2 - x) < 0.28 * y): triggered.add(17)
    if ((z - x) < 0.28 * y) != ((z * 1.5 - x) < 0.28 * y): triggered.add(18)
    if ((z - x) < 0.28 * y) != ((z - y) < 0.28 * y): triggered.add(19)
    if ((z - x) < 0.28 * y) != ((z - x) < 0.28 * x): triggered.add(20)
    if ((z - x) < 0.28 * y) != ((z - x) < 0.28 * z): triggered.add(21)
    if ((z - x) < 0.28 * y) != ((z - x) < 0.48 * y): triggered.add(22)
    if ((z - x) < 0.28 * y) != ((z - x) < 0.18 * y): triggered.add(23)
    if ((z - x) < 0.28 * y) != ((z + x) < 0.28 * y): triggered.add(24)
    if ((z - x) < 0.28 * y) != ((z - x * 1.2) < 0.28 * y): triggered.add(25)
    if ((z - x) < 0.28 * y) != ((z - x * 2) < 0.28 * y): triggered.add(26)

    # --- 分支 27-41 (原 (process_a^3 + process_b^3) < process_c^2 的变异) ---
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2.5 + y ** 3) < z ** 2): triggered.add(27)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 2.5) < z ** 2): triggered.add(28)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 - y ** 3) < z ** 2): triggered.add(29)
    if ((x ** 3 + y ** 3) < z ** 2) != ((y ** 3 + y ** 3) < z ** 2): triggered.add(30)
    if ((x ** 3 + y ** 3) < z ** 2) != ((z ** 3 + y ** 3) < z ** 2): triggered.add(31)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + x ** 3) < z ** 2): triggered.add(32)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + z ** 3) < z ** 2): triggered.add(33)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 2.5) < z ** 2): triggered.add(34)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < x ** 2): triggered.add(35)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < y ** 2): triggered.add(36)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 2.3): triggered.add(37)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 1.5): triggered.add(38)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * 3) < z ** 2): triggered.add(39)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x * 3 + y ** 3) < z ** 2): triggered.add(40)
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z * 2): triggered.add(41)

    # --- 分支 42-54 (原 abs(process_c-(process_a+process_b))<2.5 and abs(process_b-process_a*1.25)<1.5 的变异) ---
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z * 2 - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5): triggered.add(42)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + x)) < 2.5 and abs(y - x * 1.25) < 1.5): triggered.add(43)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x - y)) < 2.5 and abs(y - x * 1.25) < 1.5): triggered.add(44)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (y + y)) < 2.5 and abs(y - x * 1.25) < 1.5): triggered.add(45)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x * 0.9 + y)) < 2.5 and abs(y - x * 1.25) < 1.5): triggered.add(46)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + y * 0.8)) < 2.5 and abs(y - x * 1.25) < 1.5): triggered.add(47)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + y)) < 7 and abs(y - x * 1.25) < 1.5): triggered.add(48)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + y)) < 2.5 and abs(z - x * 1.25) < 1.5): triggered.add(49)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + y)) < 2.5 and abs(y * 1.5 - x * 1.25) < 1.5): triggered.add(50)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + y)) < 2.5 and abs(y - x * 1.75) < 1.5): triggered.add(51)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + y)) < 2.5 and abs(y - z * 1.25) < 1.5): triggered.add(52)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 8): triggered.add(53)
    if (abs(z - (x + y)) < 2.5 and abs(y - x * 1.25) < 1.5) != (abs(z - (x + y)) < 2.5 and abs(y - x ** 1.25) < 1.5): triggered.add(54)

    # --- 分支 55-67 (原 x/(y+0.01)>3.5 and y/(z+0.01)<0.3 的变异) ---
    # 注意：原代码中分支56和66都赋值67，已修正
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x * 1.3 / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3): triggered.add(55)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y * 0.6 + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3): triggered.add(56)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y + 0.01)) > 8 and (y / (z + 0.01)) < 0.3): triggered.add(57)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y + 0.01)) > 3.5 and (z / (z + 0.01)) < 0.3): triggered.add(58)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y + 0.01)) > 3.5 and (x / (z + 0.01)) < 0.3): triggered.add(59)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((y / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3): triggered.add(60)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((z / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3): triggered.add(61)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x % (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3): triggered.add(62)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y + 0.01)) > 3.5 and (y % (z + 0.01)) < 0.3): triggered.add(63)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y + 0.01)) > 3.5 or (y / (z + 0.01)) < 0.3): triggered.add(64)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.2): triggered.add(65)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 1.3): triggered.add(66)
    if ((x / (y + 0.01)) > 3.5 and (y / (z + 0.01)) < 0.3) != ((x / (y + 0.01)) > 3.5 and (y / (z * 0.7 + 0.01)) < 0.3): triggered.add(67)

    # --- 分支 68-81 (原 abs(x-y)>14 and abs(y-z)>16 and abs(x-z)<7 的变异) ---
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x * 1.2 - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7): triggered.add(68)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x * 1.6 - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7): triggered.add(69)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y * 1.2) > 14 and abs(y - z) > 16 and abs(x - z) < 7): triggered.add(70)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - z) > 14 and abs(y - z) > 16 and abs(x - z) < 7): triggered.add(71)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(x - z) > 16 and abs(x - z) < 7): triggered.add(72)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(y - x) > 16 and abs(x - z) < 7): triggered.add(73)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(y - z) > 16 and abs(y - z) < 7): triggered.add(74)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - y) < 7): triggered.add(75)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 12): triggered.add(76)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(y - z) > 16 and abs(x * 2 - z) < 7): triggered.add(77)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(y * 3 - z) > 16 and abs(x - z) < 7): triggered.add(78)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(y - z * 1.6) > 16 and abs(x - z) < 7): triggered.add(79)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 14 and abs(y - z * 0.7) > 16 and abs(x - z) < 7): triggered.add(80)
    if (abs(x - y) > 14 and abs(y - z) > 16 and abs(x - z) < 7) != (abs(x - y) > 20 and abs(y - z) > 16 and abs(x - z) < 7): triggered.add(81)

    # --- 分支 82-93 (原 (x>85 or x<8) and (y>80 or y<5) and (z>75 or z<4) 的变异) ---
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x * x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(82)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x * y > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(83)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x * z > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(84)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x > 15 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(85)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x > 85 or x < 8) and (y * x > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(86)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x > 85 or x < 8) and (y * y > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(87)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x > 85 or x < 8) and (y * z > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(88)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x > 85 or x < 8) and (y > 10 or y < 5) and (z > 75 or z < 4)): triggered.add(89)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x * 50 > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(90)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x * 80 > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(91)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x > 15 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)): triggered.add(92)
    if ((x > 85 or x < 8) and (y > 80 or y < 5) and (z > 75 or z < 4)) != ((x > 85 or x < 8) and (y > 80 or y < 5) and (z * z > 75 or z < 4)): triggered.add(93)

    # --- 分支 94-107 (原 (x+y)^1.2<z^1.8 and (x+y+z)/3>45 的变异) ---
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((y + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45): triggered.add(94)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((z + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45): triggered.add(95)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + x) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45): triggered.add(96)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + z) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45): triggered.add(97)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.8 < z ** 1.8 and (x + y + z) / 3 > 45): triggered.add(98)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.2 < x ** 1.8 and (x + y + z) / 3 > 45): triggered.add(99)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.2 < y ** 1.8 and (x + y + z) / 3 > 45): triggered.add(100)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 0.8 < z ** 1.8 and (x + y + z) / 3 > 45): triggered.add(101)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.2 < z ** 1.8 and (x + x + z) / 3 > 45): triggered.add(102)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.2 < z ** 1.8 and (x + z + z) / 3 > 45): triggered.add(103)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.2 < z ** 1.8 and (x + y + x) / 3 > 45): triggered.add(104)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.2 < z ** 1.8 and (x + y + y) / 3 > 45): triggered.add(105)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.2 < z ** 1.8 and (x + 50 + z) / 3 > 45): triggered.add(106)
    if ((x + y) ** 1.2 < z ** 1.8 and (x + y + z) / 3 > 45) != ((x + y) ** 1.2 < z ** 1.8 and (60 + y + z) / 3 > 45): triggered.add(107)

    return triggered


# === target path definitions ===
targetPaths = [
    {1, 2, 5, 6, 8, 9, 10, 14, 16, 17, 18, 24, 29, 69, 71, 72, 74, 75, 77, 79, 82, 83, 84, 85, 90, 91, 92, 98},
    {1, 2, 5, 6, 8, 9, 10, 14, 16, 17, 18, 24, 29, 71, 72, 74, 75, 77, 79, 82, 83, 84, 85, 90, 91, 92, 98},
    {1, 2, 4, 5, 6, 7, 8, 9, 10, 14, 17, 18, 24, 29, 71, 72, 74, 75, 77, 82, 83, 84, 85, 90, 91, 92, 98},
    {3, 4, 7, 12, 13, 15, 17, 18, 24, 29, 68, 71, 72, 74, 75, 77, 79, 82, 83, 84, 85, 90, 91, 92, 98},
    {1, 2, 4, 5, 6, 7, 10, 14, 15, 16, 17, 18, 20, 21, 24, 29, 77, 94, 98, 99, 102, 103, 104, 106},
    {1, 3, 4, 5, 7, 15, 30, 31, 33, 35, 36, 38, 41, 61, 62, 64, 86, 87, 88, 89, 99, 102, 104, 105},
    {4, 5, 7, 16, 26, 29, 42, 43, 44, 45, 46, 47, 49, 50, 51, 52, 54, 82, 83, 84, 85, 90, 91, 92},
    {1, 3, 4, 5, 6, 7, 8, 9, 10, 14, 16, 19, 21, 22, 25, 26, 29, 82, 83, 84, 85, 90, 91, 92, 98},
    {1, 4, 6, 7, 15, 16, 17, 18, 20, 21, 23, 24, 29, 32, 39, 93, 94, 98, 99, 102, 103, 104, 106},
    {1, 4, 5, 6, 7, 15, 28, 29, 32, 34, 37, 39, 61, 62, 64, 86, 87, 88, 89, 99, 102, 104, 105},
    {16, 20, 21, 22, 25, 26, 57, 58, 59, 60, 62, 63, 65, 67, 76, 86, 87, 88, 89, 98, 100, 105},
    {2, 16, 20, 21, 25, 26, 40, 49, 57, 58, 59, 60, 63, 65, 86, 87, 88, 89, 98, 100, 105},
    {3, 4, 5, 7, 8, 9, 14, 15, 16, 27, 28, 29, 30, 34, 37, 39, 40, 61, 62, 64, 104, 105},
    {1, 3, 4, 5, 6, 7, 8, 9, 14, 15, 28, 29, 32, 34, 37, 39, 61, 62, 64, 103, 106, 107},
    {1, 2, 5, 6, 9, 10, 14, 15, 16, 24, 29, 71, 72, 74, 75, 77, 96, 97, 99, 100, 101},
    {12, 13, 15, 17, 18, 19, 24, 70, 71, 72, 74, 75, 77, 80, 81, 86, 87, 88, 89, 98},
    {3, 12, 13, 15, 17, 18, 24, 29, 73, 78, 79, 80, 82, 83, 84, 85, 90, 91, 92, 98},
    {1, 6, 15, 16, 17, 18, 20, 21, 29, 32, 39, 93, 94, 95, 98, 99, 102, 103, 106},
    {3, 4, 7, 11, 12, 13, 15, 17, 18, 24, 29, 68, 71, 72, 73, 74, 75, 77, 81, 98},
    {2, 16, 20, 21, 22, 25, 26, 55, 56, 61, 62, 64, 76, 86, 87, 88, 89, 98},
    {4, 5, 7, 10, 14, 16, 26, 29, 46, 48, 82, 83, 84, 85, 90, 91, 92, 98},
    {3, 4, 5, 7, 10, 14, 16, 26, 29, 53, 82, 83, 84, 85, 90, 91, 92},
    {17, 18, 19, 24, 64, 66, 76, 86, 87, 88, 89, 98}
]


# 
target_paths = [set(path) for path in targetPaths]

targetPaths = [set(path) for path in target_paths]
NUM_PATHS = len(targetPaths)

def jaccard_similarity(set1, set2):
    """Compute Jaccard similarity"""
    if not set2:
        return 0.0

    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0

# === Path Similarity ===
def compute_path_similarity_matrix(paths):
    """Path Similarity"""
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix

def group_paths_by_similarity(paths, threshold_percentile=50):
    """SimilarityPath """
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.percentile(avg_sim_scores, threshold_percentile)

    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]

    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)

    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group

# === Sample generation ===
def compute_robustness(state, path, sample_size=9):
    """()"""
    base = execute_Tr(state[0], state[1], state[2])
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0

    # : 
    # weather1-6, +/-1
    # time_period1-6, +/-1
    # z1-60, +/-6(10%), +/-3(5%)
    deltas = [
        (-1, -1, -6), (0, -1, 0), (1, -1, 6),
        (-1, 0, -6), (1, 0, 6),
        (-1, 1, -6), (0, 1, 0), (1, 1, 6),
        (0, 0, 0)
    ]

    for dw, dt, dz in deltas[:sample_size]:
        if dw == dt == dz == 0:
            continue

        neighbor_weather = int(np.clip(state[0] + dw, MIN_X, MAX_X))
        neighbor_time = int(np.clip(state[1] + dt, MIN_Y, MAX_Y))
        neighbor_z = int(np.clip(state[2] + dz, MIN_Z, MAX_Z))
        neighbor = (neighbor_weather, neighbor_time, neighbor_z)

        n_trig = execute_Tr(neighbor[0], neighbor[1], neighbor[2])
        if not n_trig:
            continue

        rob += jaccard_similarity(n_trig, base)
        neighbors += 1

    return rob / neighbors if neighbors > 0 else 0.0

def generate_samples_for_all_paths(num_candidates=2000, top_k=200, run_id=1):
    """Path """
    BEST_WEIGHTS = [0.55, 0.25, 0.2]

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_individual.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Individual Path {path_id} (Weighted Screening) - Run {run_id}\n")
            f.write("weather time_period z\tScore\tSimilarity\tRobustness\tLengthDiff\n")
            for s in samples:
                weather, time_period, z = s['state']
                f.write(
                    f"{weather} {time_period} {z}\t{s['score']:.4f}\t{s['similarity']:.4f}\t"
                    f"{s['robustness']:.4f}\t{s['length_diff']:.4f}\n"
                )

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_individual"

    for path_idx in range(len(targetPaths)):
        path = targetPaths[path_idx]
        candidate_samples = []
        attempts = 0
        max_attempts = num_candidates * 10

        while len(candidate_samples) < num_candidates and attempts < max_attempts:
            attempts += 1

            weather = np.random.randint(MIN_X, MAX_X + 1)
            time_period = np.random.randint(MIN_Y, MAX_Y + 1)
            z = np.random.randint(MIN_Z, MAX_Z + 1)
            state = (weather, time_period, z)

            triggered = execute_Tr(weather, time_period, z)
            if not triggered:
                continue

            sim = jaccard_similarity(triggered, path)
            rob = compute_robustness(state, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))

            candidate_samples.append({
                'state': state,
                'similarity': sim,
                'robustness': rob,
                'length_diff': len_diff,
                'triggered': triggered
            })

        if candidate_samples:
            for sample in candidate_samples:
                score = (BEST_WEIGHTS[0] * sample['similarity'] +
                         BEST_WEIGHTS[1] * sample['robustness'] +
                         BEST_WEIGHTS[2] * sample['length_diff'])
                sample['score'] = score

            candidate_samples.sort(key=lambda x: x['score'], reverse=True)
            selected_samples = candidate_samples[:top_k]
            save_samples(path_id=path_idx + 1, samples=selected_samples, base_dir=base_dir)

# === (: )===
class SharedExperienceReplay:
    """()"""

    def __init__(self, capacity=20000):
        self.capacity = capacity
        self.buffer = deque(maxlen=self.capacity)
        self.priorities = deque(maxlen=self.capacity)

    def append(self, experience):
        """"""
        self.buffer.append(experience)
        self.priorities.append(experience[-1])

    def sample(self, batch_size, alpha=0.6):
        """"""
        if len(self.buffer) < batch_size:
            return [], [], []

        priorities = np.array(self.priorities, dtype=np.float64)
        priorities = np.power(priorities, alpha)
        probabilities = priorities / np.sum(priorities)

        batch_indices = np.random.choice(len(self.buffer), batch_size, p=probabilities, replace=False)
        batch = [self.buffer[idx] for idx in batch_indices]

        return batch, batch_indices, probabilities[batch_indices]

    def update_priorities(self, batch_indices, td_errors):
        """"""
        for idx, td_error in zip(batch_indices, td_errors):
            if idx < len(self.priorities):
                self.priorities[idx] = max(abs(td_error), 1e-6)

    def __len__(self):
        return len(self.buffer)

    def get_high_reward_samples(self, target_path, num_samples=20):
        """(, )"""
        if len(self.buffer) == 0:
            return []

        samples_with_scores = []
        seen_states = set()  # 

        for experience in self.buffer:
            state_tensor = experience[0]
            state_norm = state_tensor.cpu().numpy().flatten()

            # 
            state_tuple = denormalize_state((state_norm[0], state_norm[1], state_norm[2]))

            # 
            if state_tuple in seen_states:
                continue
            seen_states.add(state_tuple)

            triggered = execute_Tr(state_tuple[0], state_tuple[1], state_tuple[2])
            reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)

            samples_with_scores.append((state_tuple, reward, sim, triggered))

        samples_with_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_scores[:num_samples]

def load_path_data(file_path):
    """screeningPath """
    path_data = []

    if not os.path.exists(file_path):
        print(f":  {file_path}")
        return path_data

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
            for line in lines[2:]:
                parts = line.strip().split("\t")
                if parts:
                    values = parts[0].split()
                    if len(values) >= 3:
                        state = (int(values[0]), int(values[1]), int(values[2]))
                        path_data.append(state)
    except Exception as e:
        print(f" {file_path}: {e}")

    return path_data

# === DQN ===
class DQN(nn.Module):
    """Q"""

    def __init__(self, state_dim, action_dim, hidden_dims=[128, 64]):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, hidden_dims[0])
        self.fc2 = nn.Linear(hidden_dims[0], hidden_dims[1])
        self.fc3 = nn.Linear(hidden_dims[1], action_dim)
        self.dropout = nn.Dropout(0.1)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = self.dropout(x)
        x = torch.relu(self.fc2(x))
        return self.fc3(x)

# === DQN Agent with PER(: )===
class DQNAgentWithPER:
    """DQN()"""

    def __init__(self, state_dim, action_dim, replay_buffer,
                 gamma=0.99, epsilon=1.0, epsilon_decay=0.995,
                 epsilon_min=0.1, learning_rate=0.001, alpha=0.6, beta=0.4):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer
        self.alpha = alpha
        self.beta = beta

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        """
        ()
        30, : 
        - weather: +/-1, 0(x2)
        - time_period: +/-1, 0(x2)
        - z: +/-12(20%), +/-6(10%), +/-3(5%), 0(x2)
        """
        delta_values_weather_time = [1, 0, 0, -1]  # 
        delta_values_z = [12, 6, 3, 0, 0, -3, -6, -12]  # z

        dim = action_idx // 10
        delta_idx = action_idx % 10

        if dim == 0:  # weather
            if delta_idx >= 4:
                delta_idx = 3
            return (delta_values_weather_time[delta_idx], 0, 0)
        elif dim == 1:  # time_period
            if delta_idx >= 4:
                delta_idx = 3
            return (0, delta_values_weather_time[delta_idx], 0)
        elif dim == 2:  # z
            if delta_idx >= 8:
                delta_idx = 7
            return (0, 0, delta_values_z[delta_idx])

    def act(self, state_norm, legal_actions=None):
        """()"""
        if legal_actions is None:
            legal_actions = list(range(self.action_dim))

        if not legal_actions:
            return None

        if random.random() < self.epsilon:
            return random.choice(legal_actions)

        state_tensor = torch.tensor(state_norm, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)[0]

        legal_q_values = q_values[legal_actions]
        best_legal_idx = torch.argmax(legal_q_values).item()
        return legal_actions[best_legal_idx]

    def get_legal_actions(self, state):
        """()"""
        legal_actions = []

        for action_idx in range(self.action_dim):
            dw, dt, dz = self.decode_action(action_idx)

            next_weather = state[0] + dw
            next_time = state[1] + dt
            next_z = state[2] + dz

            if (MIN_X <= next_weather <= MAX_X and
                    MIN_Y <= next_time <= MAX_Y and
                    MIN_Z <= next_z <= MAX_Z):
                legal_actions.append(action_idx)

        return legal_actions

    def store_transition(self, state, action, reward, next_state, done):
        """()"""
        # 
        state_norm = normalize_state(state)
        next_state_norm = normalize_state(next_state)

        state_tensor = torch.tensor(state_norm, dtype=torch.float32).unsqueeze(0).to(device)
        next_state_tensor = torch.tensor(next_state_norm, dtype=torch.float32).unsqueeze(0).to(device)

        with torch.no_grad():
            q_values = self.model(state_tensor)
            next_q_values = self.target_model(next_state_tensor)
            max_next_q = next_q_values.max(1)[0]
            target_q = reward + (self.gamma * max_next_q * (1 - done))
            td_error = abs(q_values[0][action].item() - target_q.item())

        self.replay_buffer.append((state_tensor, action, reward, next_state_tensor, done, td_error))
        return td_error

    def train(self, batch_size=32):
        """"""
        if len(self.replay_buffer) < batch_size:
            return 0.0

        batch, batch_indices, probabilities = self.replay_buffer.sample(batch_size, alpha=self.alpha)

        if not batch:
            return 0.0

        states, actions, rewards, next_states, dones, _ = zip(*batch)

        weights = (len(self.replay_buffer) * probabilities) ** (-self.beta)
        weights = weights / weights.max()
        weights = torch.tensor(weights, dtype=torch.float32).to(device)

        states = torch.cat(states).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.cat(next_states).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q = self.target_model(next_states).max(1)[0].detach()
        target_q = rewards + (self.gamma * next_max_q * (1 - dones))

        td_errors = current_q - target_q
        weighted_loss = (td_errors.pow(2) * weights).mean()

        self.optimizer.zero_grad()
        weighted_loss.backward()
        torch.nn.utils.clip_grad_norm_(self.model.parameters(), 1.0)
        self.optimizer.step()

        new_priorities = abs(td_errors.detach().cpu().numpy())
        self.replay_buffer.update_priorities(batch_indices, new_priorities)

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

        return weighted_loss.item()

    def update_target_model(self):
        """"""
        self.target_model.load_state_dict(self.model.state_dict())

# ===  ===
def generate_and_train_for_individual_paths(path_documents, repeats=5, batch_size=32, run_id=1):
    """"""
    state_dim = 3
    action_dim = 30

    shared_replay_buffer = SharedExperienceReplay(capacity=20000)
    agent = DQNAgentWithPER(state_dim, action_dim, shared_replay_buffer)

    total_cumulative_reward = 0
    path_rewards = {}

    print(f"\n===  {run_id}/20 Start training(+)===")
    start_time = time.time()

    SAMPLES_PER_BATCH = 50
    NUM_BATCHES = 4
    STEPS_PER_SAMPLE = 3

    for path_idx in range(len(targetPaths)):
        path_id = path_idx + 1
        print(f"\n{'=' * 60}")
        print(f"Path  {path_id}/{NUM_PATHS}")
        print(f"{'=' * 60}")

        file_path = os.path.join(path_documents, f"path{path_id}_individual.txt")
        if not os.path.exists(file_path):
            print(f"  :  {file_path}")
            continue

        path_data = load_path_data(file_path)
        if not path_data:
            print(f"  : Path  {path_id} ")
            continue

        target_path = targetPaths[path_idx]

        if path_idx not in path_rewards:
            path_rewards[path_idx] = 0

        for repeat_idx in range(repeats):
            print(f"\n  Run  {repeat_idx + 1}/{repeats} ")

            for batch_idx in range(NUM_BATCHES):
                batch_start = batch_idx * SAMPLES_PER_BATCH
                batch_end = min(batch_start + SAMPLES_PER_BATCH, len(path_data))

                print(f"     {batch_idx + 1}/{NUM_BATCHES} ( {batch_start}-{batch_end})")

                for sample_idx in range(batch_start, batch_end):
                    state = path_data[sample_idx]
                    prev_state = None
                    prev_triggered = None

                    for step in range(STEPS_PER_SAMPLE):
                        legal_actions = agent.get_legal_actions(state)

                        if not legal_actions:
                            break

                        # 
                        state_norm = normalize_state(state)
                        action = agent.act(state_norm, legal_actions)
                        if action is None:
                            break

                        dw, dt, dz = agent.decode_action(action)

                        next_state = (
                            int(np.clip(state[0] + dw, MIN_X, MAX_X)),
                            int(np.clip(state[1] + dt, MIN_Y, MAX_Y)),
                            int(np.clip(state[2] + dz, MIN_Z, MAX_Z))
                        )

                        triggered = execute_Tr(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered,
                                                prev_triggered, prev_state)
                        done = (step == STEPS_PER_SAMPLE - 1)

                        # 
                        agent.store_transition(state, action, reward, next_state, done)

                        prev_state = state
                        prev_triggered = triggered
                        state = next_state

                        total_cumulative_reward += reward
                        path_rewards[path_idx] += reward

                if len(agent.replay_buffer) >= batch_size:
                    loss = agent.train(batch_size)
                    print(f"       {batch_idx + 1} completed, : {loss:.4f}")

                if (batch_idx + 1) % 2 == 0:
                    agent.update_target_model()
                    print(f"      completed {batch_idx + 1} , ")

        print(f"\nPath  {path_id} completed, : {path_rewards[path_idx]:.2f}")
        print(f"Shared Buffer Size: {len(shared_replay_buffer)}")

    training_time = time.time() - start_time
    print(f"\n===  {run_id}/20 completed, : {training_time:.2f} seconds ===")

    return agent, shared_replay_buffer, total_cumulative_reward, path_rewards, training_time

# === Excel ===
def create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir):
    """Excel"""
    os.makedirs(output_dir, exist_ok=True)

    similar_group_paths = [idx + 1 for idx in similar_group]
    isolated_group_paths = [idx + 1 for idx in isolated_group]

    wb = Workbook()

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    header_color = "4472C4"
    similar_group_color = "E2EFDA"
    isolated_group_color = "FCE4D6"
    stats_color = "FFF2CC"

    # === 1: Path  ===
    ws_paths = wb.active
    ws_paths.title = "Path "

    path_headers = ['Path ID', ''] + [f'Run {i}' for i in range(1, 21)] + \
                   ['Average Similarity', 'Maximum Similarity', 'Minimum Similarity', 'Standard deviation']

    for col, header in enumerate(path_headers, 1):
        cell = ws_paths.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_paths.row_dimensions[1].height = 30

    for path_id in range(1, NUM_PATHS + 1):
        row = path_id + 1

        if path_id in similar_group_paths:
            group_type = "High-correlation path group"
            row_color = similar_group_color
        elif path_id in isolated_group_paths:
            group_type = "Low-correlation path group"
            row_color = isolated_group_color
        else:
            group_type = "Ungrouped"
            row_color = "FFFFFF"

        cell = ws_paths.cell(row=row, column=1, value=f"Path {path_id}")
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        cell.border = thin_border

        cell = ws_paths.cell(row=row, column=2, value=group_type)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        cell.border = thin_border

        path_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            sim = run_data['path_similarities'].get(path_id, {}).get('avg_similarity', 0.0)
            path_similarities.append(sim)

            cell = ws_paths.cell(row=row, column=3 + run_idx, value=round(sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        stats_values = [
            round(np.mean(path_similarities), 4),
            round(np.max(path_similarities), 4),
            round(np.min(path_similarities), 4),
            round(np.std(path_similarities), 4)
        ]

        for i, value in enumerate(stats_values):
            cell = ws_paths.cell(row=row, column=23 + i, value=value)
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
            cell.font = Font(bold=True, size=10)
            cell.border = thin_border

    ws_paths.column_dimensions['A'].width = 13
    ws_paths.column_dimensions['B'].width = 16
    for col in range(3, 23):
        ws_paths.column_dimensions[get_column_letter(col)].width = 10
    for col in range(23, 27):
        ws_paths.column_dimensions[get_column_letter(col)].width = 13

    # === 2:  ===
    ws_groups = wb.create_sheet("")

    group_headers = ['Group Name', 'Included Paths'] + [f'Run {i}' for i in range(1, 21)] + \
                    ['Average Similarity', 'Standard deviation']

    for col, header in enumerate(group_headers, 1):
        cell = ws_groups.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_groups.row_dimensions[1].height = 30

    row = 2

    # High-correlation path group
    cell = ws_groups.cell(row=row, column=1, value="High-correlation path group")
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    cell.border = thin_border

    cell = ws_groups.cell(row=row, column=2, value=','.join(map(str, similar_group_paths)))
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    cell.border = thin_border

    group_similarities = []
    for run_idx, run_data in enumerate(all_runs_data):
        group_sim = np.mean([
            run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
            for p in similar_group_paths
        ])
        group_similarities.append(group_sim)

        cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(group_sim, 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    cell = ws_groups.cell(row=row, column=23, value=round(np.mean(group_similarities), 4))
    cell.number_format = '0.0000'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border

    cell = ws_groups.cell(row=row, column=24, value=round(np.std(group_similarities), 4))
    cell.number_format = '0.0000'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border

    row += 1

    # Low-correlation path group
    if isolated_group_paths:
        cell = ws_groups.cell(row=row, column=1, value="Low-correlation path group")
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        cell.border = thin_border

        cell = ws_groups.cell(row=row, column=2, value=','.join(map(str, isolated_group_paths)))
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        cell.border = thin_border

        isolated_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            iso_sim = np.mean([
                run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
                for p in isolated_group_paths
            ])
            isolated_similarities.append(iso_sim)

            cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(iso_sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        cell = ws_groups.cell(row=row, column=23, value=round(np.mean(isolated_similarities), 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border

        cell = ws_groups.cell(row=row, column=24, value=round(np.std(isolated_similarities), 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border

    ws_groups.column_dimensions['A'].width = 16
    ws_groups.column_dimensions['B'].width = 22
    for col in range(3, 23):
        ws_groups.column_dimensions[get_column_letter(col)].width = 10
    ws_groups.column_dimensions[get_column_letter(23)].width = 14
    ws_groups.column_dimensions[get_column_letter(24)].width = 12

    # === 3: Detailed Sample Data ===
    ws_samples = wb.create_sheet("Detailed Sample Data")

    sample_headers = ['Run', 'Path ID', 'Sample ID', '', '',
                      '', 'Similarity', 'Triggered Rule Set']

    for col, header in enumerate(sample_headers, 1):
        cell = ws_samples.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_samples.row_dimensions[1].height = 30

    sample_row = 2
    for run_idx, run_data in enumerate(all_runs_data, 1):
        for path_id in range(1, NUM_PATHS + 1):
            samples = run_data['path_samples'].get(path_id, [])

            if path_id in similar_group_paths:
                path_color = similar_group_color
            elif path_id in isolated_group_paths:
                path_color = isolated_group_color
            else:
                path_color = "FFFFFF"

            for sample_idx, (state_tuple, reward, sim, triggered) in enumerate(samples, 1):
                weather, time_period, z = state_tuple
                triggered_str = ','.join(map(str, sorted(triggered)))

                cell = ws_samples.cell(row=sample_row, column=1, value=f"Run {run_idx}")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=path_color, end_color=path_color, fill_type="solid")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=2, value=f"Path {path_id}")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=path_color, end_color=path_color, fill_type="solid")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=3, value=sample_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                for col_offset, value in enumerate([weather, time_period, z]):
                    cell = ws_samples.cell(row=sample_row, column=4 + col_offset, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=7, value=round(sim, 4))
                cell.number_format = '0.0000'
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=8, value=f"{{{triggered_str}}}")
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border

                sample_row += 1

    sample_widths = [13, 13, 11, 10, 10, 12, 12, 45]
    for i, width in enumerate(sample_widths, 1):
        ws_samples.column_dimensions[get_column_letter(i)].width = width

    # === 4: Run Statistics Summary ===
    ws_summary = wb.create_sheet("Run Statistics Summary")

    summary_headers = ['Run', 'Training Time( seconds)', 'Overall Average Similarity', 'Maximum Similarity',
                       'Minimum Similarity', '', '', 'Shared Buffer Size']

    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_summary.row_dimensions[1].height = 30

    for run_idx, run_data in enumerate(all_runs_data, 1):
        row = run_idx + 1

        # Average Similarity
        high_group_avg = np.mean([
            run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
            for p in similar_group_paths
        ])

        low_group_avg = 0.0
        if isolated_group_paths:
            low_group_avg = np.mean([
                run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
                for p in isolated_group_paths
            ])

        values = [
            f"Run {run_idx}",
            round(run_data['training_time'], 2),
            round(run_data['overall_avg_similarity'], 4),
            round(run_data['max_similarity'], 4),
            round(run_data['min_similarity'], 4),
            round(high_group_avg, 4),
            round(low_group_avg, 4),
            20000  # 
        ]

        for col, value in enumerate(values, 1):
            cell = ws_summary.cell(row=row, column=col, value=value)
            if col == 1:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 2:
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col >= 3 and col <= 7:
                cell.number_format = '0.0000'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # 
    stat_row = len(all_runs_data) + 2
    stat_labels = ['', '/', '', '', '', '', '', '']

    for col, label in enumerate(stat_labels, 1):
        cell = ws_summary.cell(row=stat_row, column=col, value=label)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    stat_row += 1

    # 
    training_times = [r['training_time'] for r in all_runs_data]
    overall_avgs = [r['overall_avg_similarity'] for r in all_runs_data]
    max_sims = [r['max_similarity'] for r in all_runs_data]
    min_sims = [r['min_similarity'] for r in all_runs_data]

    high_group_avgs = []
    low_group_avgs = []
    for run_data in all_runs_data:
        high_avg = np.mean([
            run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
            for p in similar_group_paths
        ])
        high_group_avgs.append(high_avg)

        if isolated_group_paths:
            low_avg = np.mean([
                run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
                for p in isolated_group_paths
            ])
            low_group_avgs.append(low_avg)

    stat_values = [
        '',
        round(np.sum(training_times), 2),
        round(np.mean(overall_avgs), 4),
        round(np.max(max_sims), 4),
        round(np.min(min_sims), 4),
        round(np.mean(high_group_avgs), 4),
        round(np.mean(low_group_avgs), 4) if low_group_avgs else 0.0,
        20000
    ]

    for col, value in enumerate(stat_values, 1):
        cell = ws_summary.cell(row=stat_row, column=col, value=value)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        if col == 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col == 2:
            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col >= 3 and col <= 7:
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    summary_widths = [13, 16, 18, 14, 14, 16, 16, 14]
    for i, width in enumerate(summary_widths, 1):
        ws_summary.column_dimensions[get_column_letter(i)].width = width

    output_path = os.path.join(output_dir, "20 run_.xlsx")
    wb.save(output_path)
    print(f"\n Consolidated Excel report generated: {output_path}")
    print(f"   4: Path , , Detailed Sample Data, Run Statistics Summary")

def run_20_times_training():
    """20"""
    model_path_base = r"D:\Experiment\CNN\DQNNEW\saved_models_new_vars"
    path_documents = r"D:\Experiment\CNN\DQNNEW\path_samples_individual"
    output_dir = r"D:\Experiment\CNN\ComparisonExperiment2\excel_reports_new_vars"

    os.makedirs(model_path_base, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    similar_group, isolated_group = group_paths_by_similarity(targetPaths)

    similar_group_display = [idx + 1 for idx in similar_group]
    isolated_group_display = [idx + 1 for idx in isolated_group]

    print("=" * 60)
    print("20 - ")
    print("=" * 60)
    print(f"\nPath : {NUM_PATHS} ")
    print(f"\nAutomatic grouping results:")
    print(f"  Similar path group: {similar_group_display}")
    print(f"  Isolated path group: {isolated_group_display}")
    print(f"\n:")
    print(f"   (weather): {MIN_X}-{MAX_X}")
    print(f"   (time_period): {MIN_Y}-{MAX_Y}")
    print(f"   (z): {MIN_Z}-{MAX_Z}")
    print(f"\n:")
    print(f"  x = z ()")
    print(f"  y = (weather * time_period * 10 + z) % 100 + 1")
    print(f"\n:")
    print(f"  : [0, 1]")
    print(f"  : ")
    print(f"  : (seen_states)")
    print(f"\n (30):")
    print(f"  weather: +/-1, 0(x2)")
    print(f"  time_period: +/-1, 0(x2)")
    print(f"  z: +/-12, +/-6, +/-3, 0(x2)")
    print("\n" + "=" * 60)

    all_runs_data = []
    total_start_time = time.time()

    for run_id in range(1, 21):
        print(f"\n{'=' * 60}")
        print(f"Start run  {run_id}/20  run")
        print(f"{'=' * 60}")

        print(f"[Run {run_id}] Generating samples...")
        generate_samples_for_all_paths(num_candidates=2000, top_k=200, run_id=run_id)

        print(f"[Run {run_id}] Start training...")
        agent, shared_buffer, total_cumulative_reward, path_rewards, training_time = \
            generate_and_train_for_individual_paths(path_documents, repeats=5,
                                                    batch_size=32, run_id=run_id)

        model_path = os.path.join(model_path_base, f"trained_model_run_{run_id}.pth")
        torch.save({
            'model_state_dict': agent.model.state_dict(),
            'optimizer_state_dict': agent.optimizer.state_dict(),
            'epsilon': agent.epsilon,
            'run_id': run_id,
            'normalized': True,
            'value_ranges': {
                'weather': [MIN_X, MAX_X],
                'time_period': [MIN_Y, MAX_Y],
                'z': [MIN_Z, MAX_Z]
            }
        }, model_path)
        print(f"[Run {run_id}] Model saved: {model_path}")

        run_data = {
            'run_id': run_id,
            'training_time': training_time,
            'total_reward': total_cumulative_reward,
            'path_rewards': path_rewards,
            'path_similarities': {},
            'path_samples': {}
        }

        all_similarities = []
        for path_idx in range(len(targetPaths)):
            target_path = targetPaths[path_idx]
            high_reward_samples = shared_buffer.get_high_reward_samples(target_path, num_samples=20)

            if high_reward_samples:
                similarities = [sim for _, _, sim, _ in high_reward_samples]
                run_data['path_similarities'][path_idx + 1] = {
                    'avg_similarity': np.mean(similarities),
                    'max_similarity': np.max(similarities),
                    'min_similarity': np.min(similarities),
                    'sample_count': len(similarities)
                }
                run_data['path_samples'][path_idx + 1] = high_reward_samples
                all_similarities.extend(similarities)
            else:
                run_data['path_similarities'][path_idx + 1] = {
                    'avg_similarity': 0.0,
                    'max_similarity': 0.0,
                    'min_similarity': 0.0,
                    'sample_count': 0
                }
                run_data['path_samples'][path_idx + 1] = []

        if all_similarities:
            run_data['overall_avg_similarity'] = np.mean(all_similarities)
            run_data['max_similarity'] = np.max(all_similarities)
            run_data['min_similarity'] = np.min(all_similarities)
        else:
            run_data['overall_avg_similarity'] = 0.0
            run_data['max_similarity'] = 0.0
            run_data['min_similarity'] = 0.0

        all_runs_data.append(run_data)

        print(f"[Run {run_id}] completed! Overall Average Similarity: {run_data['overall_avg_similarity']:.4f}")
        print(f"{'=' * 60}\n")

    total_time = time.time() - total_start_time

    print("\nGenerating consolidated Excel report...")
    create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir)

    print("\n" + "=" * 60)
    print(f"\nAll results have been saved to: {output_dir}")
    print("=" * 60)

if __name__ == "__main__":
    run_20_times_training()