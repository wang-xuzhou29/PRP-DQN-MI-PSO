import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """安全除法"""
    if denominator == 0:
        return default
    return numerator / denominator


def execute_Tr(x, y, z):
    # 初始化分支覆盖数组
    b = [0] * 99  # 根据分支数量调整大小

    if ((x * y) / (z + 1) > 150) != ((x * y) / (z + 1) > 200): b[0] = 1
    if ((x * y) / (z + 1) > 150) != ((x * y) / (z * 2 + 1) > 150): b[1] = 2
    if ((x * y) / (z + 1) > 150) != ((x * x) / (z + 1) > 150): b[2] = 3
    if ((x * y) / (z + 1) > 150) != ((x * 2 * y) / (z + 1) > 150): b[3] = 4
    if ((x * y) / (z + 1) > 150) != ((y * y) / (z + 1) > 150): b[4] = 5
    if ((x * y) / (z + 1) > 150) != ((x * y) / (z + 1) > 500): b[5] = 6
    if ((x * y) / (z + 1) > 150) != ((x * 0.5 * y) / (z + 1) > 150): b[6] = 7
    if ((x * y) / (z + 1) > 150) != ((x * y) / (z + 10) > 150): b[7] = 8
    if ((x * y) / (z + 1) > 150) != ((x * y) / (z * z + 1) > 150): b[8] = 9
    if ((x * y) / (z + 1) > 150) != ((x / y) / (z + 1) > 150): b[9] = 10

    # 验证规则2：相对偏差检测
    if ((y - x) < 0.2 * z) != ((y - x * 2) < 0.2 * z): b[10] = 11
    if ((y - x) < 0.2 * z) != ((y - x) < 0.1 * z): b[11] = 12
    if ((y - x) < 0.2 * z) != ((y - x) < 0.3 * z): b[12] = 13
    if ((y - x) < 0.2 * z) != ((y - x) < 0.5 * z): b[13] = 14
    if ((y - x) < 0.2 * z) != ((y - x) < 0.38 * z): b[14] = 15
    if ((y - x) < 0.2 * z) != ((y - x) < 0.2 * z * x): b[15] = 16
    if ((y - x) < 0.2 * z) != ((y * 1.3 - x) < 0.2 * z): b[16] = 17
    if ((y - x) < 0.2 * z) != ((y - x) < 0.2 * x): b[17] = 18
    if ((y - x) < 0.2 * z) != ((y - x) < 0.2 * y): b[18] = 19
    if ((y - x) < 0.2 * z) != ((y * 2 - x) < 0.2 * z): b[19] = 20

    # 验证规则3：立方根关系验证
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2 + y ** 3) < z ** 2): b[20] = 21
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 2) < z ** 2): b[21] = 22
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 1) < z ** 2): b[22] = 23
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 2.9): b[23] = 24
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 1.8 + y ** 3) < z ** 2): b[24] = 25
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 1 + y ** 3) < z ** 2): b[25] = 26
    if ((x ** 3 + y ** 3) < z ** 2) != ((x * 3 + y ** 3) < z ** 2): b[26] = 27
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * 3) < z ** 2): b[27] = 28
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 3): b[28] = 29
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3.2) < z ** 3): b[29] = 30

    # 验证规则6：整数同余检查
    if (int(x) % 3 == int(y) % 3 == int(z) % 3 == 0) != (
            int(x) % 2 == int(y) % 3 == int(z) % 3 == 0): b[30] = 31
    if (int(x) % 3 == int(y) % 3 == int(z) % 3 == 0) != (
            int(x) % 3 == int(y) % 2 == int(z) % 3 == 0): b[31] = 32
    if (int(x) % 3 == int(y) % 3 == int(z) % 3 == 0) != (
            int(x) % 3 == int(y) % 3 == int(z) % 2 == 0): b[32] = 33
    if (int(x) % 3 == int(y) % 3 == int(z) % 3 == 0) != (
            int(x) % 5 == int(y) % 3 == int(z) % 3 == 0): b[33] = 34
    if (int(x) % 3 == int(y) % 3 == int(z) % 3 == 0) != (
            int(x) % 3 == int(y) % 5 == int(z) % 3 == 0): b[34] = 35
    if (int(x) % 3 == int(y) % 3 == int(z) % 3 == 0) != (
            int(x) % 3 == int(y) % 3 == int(z) % 5 == 0): b[35] = 36

    # 验证规则7：比值范围检查
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (z + 0.1)) > 3 and (y / (z + 0.1)) < 0.3): b[36] = 37
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 1 and (y / (z + 0.1)) < 0.3): b[37] = 38
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (x + 0.1)) < 0.3): b[38] = 39
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (z * 1.2 + 0.1)) < 0.3): b[39] = 40
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.5): b[40] = 41
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (z * 0.1)) < 0.3): b[41] = 42
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y * 0.1)) > 3 and (y / (z + 0.1)) < 0.3): b[42] = 43
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x + (y + 0.1)) > 3 and (y / (x + 0.1)) < 0.3): b[43] = 44
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (z * 1.5 + 0.1)) < 0.3): b[44] = 45
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (z + 0)) < 0.3): b[45] = 46
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 or (y / (z + 0.1)) < 0.3): b[46] = 47
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 2)) > 3 and (y / (z + 0.1)) < 0.3): b[47] = 48
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (z + 5)) < 0.3): b[48] = 49

    # 验证规则8：差值阈值检查
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x * x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8): b[49] = 50
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 20 and abs(y - z) > 20 and abs(x - z) < 8): b[50] = 51
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 30 and abs(y - z) > 20 and abs(x - z) < 8): b[51] = 52
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y - z) > 40 and abs(x - z) < 8): b[52] = 53
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y * z - z) > 20 and abs(x - z) < 8): b[53] = 54
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 12): b[54] = 55
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z * 2) < 8): b[55] = 56
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y - z) > 20 and abs(x * 2 - z) < 8): b[56] = 57
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y * 2 - z) > 20 and abs(x - z) < 8): b[57] = 58
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x * 2 - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8): b[58] = 59
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y - z * z) > 20 and abs(x - z) < 8): b[59] = 60
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y * y - z) > 20 and abs(x - z) < 8): b[60] = 61
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y - z) > 20 and abs(x * x - z) < 8): b[61] = 62
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y * x - z) > 20 and abs(x - z) < 8): b[62] = 63

    # 验证规则9：极值范围检查
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x * 2 > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)): b[63] = 64
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 60 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)): b[64] = 65
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 115 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)): b[65] = 66
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 18) and (y > 85 or y < 2) and (z > 180 or z < 40)): b[66] = 67
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 5) and (y > 85 or y < 2) and (z > 180 or z < 40)): b[67] = 68
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 8) and (y > 130 or y < 2) and (z > 180 or z < 40)): b[68] = 69
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 8) and (y > 85 or y < 2) and (z * z > 180 or z < 40)): b[69] = 70
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 30)): b[70] = 71
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 8) and (y > 85 or y < 2) and (z * 50 > 180 or z < 40)): b[71] = 72
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 60)): b[72] = 73
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 8) and (y > 100 or y < 2) and (z > 180 or z < 40)): b[73] = 74

    # 额外的复杂验证逻辑
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.3 + y ** 0.5 > z and x * y > z ** 1.5): b[74] = 75
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.6 + y ** 0.5 > z and x * y > z ** 1.5): b[75] = 76
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.7 > z and x * y > z ** 1.5): b[76] = 77
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z and x * 0.5 > z ** 1.5): b[77] = 78
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            (x ** 0.5) * 2 + y ** 0.5 > z and x * y > z ** 1.5): b[78] = 79
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y * 0.5 > z and x * y > z ** 1.5): b[79] = 80
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z * 2 and x * y > z ** 1.5): b[80] = 81
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z * 0.5 and x * y > z ** 1.5): b[81] = 82
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z and 0.3 * y > z ** 1.5): b[82] = 83
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z and x * 0.1 > z ** 1.5): b[83] = 84
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z and 0.2 * y > z ** 1.5): b[84] = 85
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z and 0.5 * y > z ** 1.5): b[85] = 86
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z and x * y > z ** 8): b[86] = 87

    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 3 < z ** 2 * 4 and x > y): b[87] = 88
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * 3 and x > y): b[88] = 89
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * 2 and x > y): b[89] = 90
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * 4 and x * x > y): b[90] = 91
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * 4 and x * 2 > y): b[91] = 92
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + 50) ** 2 < z ** 2 * 4 and x > y): b[92] = 93
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 3 * 4 and x > y): b[93] = 94
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 3 < z ** 2 * 4 and x > y): b[94] = 95
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 4 < z ** 2 * 4 and x > y): b[95] = 96
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * 1 and x > y): b[96] = 97
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * x and x > y): b[97] = 98
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * y and x > y): b[98] = 99

    # 返回触发的分支索引集合
    triggered = set()
    for i, val in enumerate(b):
        if val > 0:
            triggered.add(val)
    return triggered


def execute_validation_rules(dx: int, dy: int, dz: int) -> Set[int]:
    """执行验证规则，返回触发的分支集合（调用105分支版本）"""
    return execute_Tr(dx, dy, dz)


def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """计算适应度"""
    generated_path = execute_validation_rules(int(particle[0]), int(particle[1]), int(particle[2]))

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """基本PSO优化器"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # 范围: x(1,200), y(2,150), z(1,200)
        self.bounds = bounds if bounds else [(1, 200), (1, 200), (2, 150)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """初始化粒子"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """更新速度和位置"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """对目标路径进行PSO优化"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(
                            int(particles[i][0]), int(particles[i][1]), int(particles[i][2])
                        ),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(
                int(gbest_particle[0]), int(gbest_particle[1]), int(gbest_particle[2])
            ),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """对所有路径运行PSO"""

    print(f"\n{'=' * 70}")
    print(f"Baseline PSO - 路径优化")
    print(f"{'=' * 70}")
    print(f"粒子数: {n_particles}, 最大迭代次数: {max_iterations}")
    print(f"路径数量: {len(target_paths)}")
    print(f"范围: x(2,100), y(1,105), z(1,110)")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"路径 {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "完美求解" if result['success'] else f"(适应度: {result['best_fitness']:.3f})"
        print(f"{status} | 耗时: {result['time']:.2f}s | 迭代次数: {result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f"求解成功: {success_count}/{len(target_paths)} ({success_rate:.1f}%) | 总耗时: {total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """多次运行实验"""

    print(f"\n{'=' * 70}")
    print(f"Baseline PSO - {num_runs}次运行")
    print(f"{'=' * 70}")
    print(f"粒子数: {n_particles}, 最大迭代次数: {max_iterations}, 路径数: {len(target_paths)}")
    print(f"范围: x(2,100), y(1,105), z(1,110)")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"--- 第 {run_idx}/{num_runs} 次运行 ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f"求解成功: {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs}次运行完成 | 总耗时: {total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """导出Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # Sheet 1: 运行汇总
    ws1 = wb.active
    ws1.title = "运行汇总"
    ws1.sheet_view.showGridLines = False

    headers = ["运行", "成功率", "求解/总数", "平均适应度", "平均迭代次数", "耗时(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f"运行 {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # Sheet 2: 路径统计
    ws2 = wb.create_sheet(title="路径统计")
    ws2.sheet_view.showGridLines = False

    headers2 = ["路径ID", "求解/总数", "成功率", "平均适应度", "平均迭代次数", "最小迭代次数", "最大迭代次数"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"路径 {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # Sheet 3: 详细结果
    ws3 = wb.create_sheet(title="详细结果")
    ws3.sheet_view.showGridLines = False

    headers3 = ["路径ID", "运行", "位置(x,y,z)", "适应度", "迭代次数", "触发路径"]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"Run {run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # Sheet 4: 目标路径
    ws4 = wb.create_sheet(title="目标路径")
    ws4.sheet_view.showGridLines = False

    headers4 = ["路径ID", "目标路径", "分支数"]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"路径 {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f"Excel已保存: {filename}")
    print(f"{'=' * 70}")
    print(f"包含工作表:")
    print(f"  1. 运行汇总      - {len(all_results)}次运行汇总")
    print(f"  2. 路径统计      - 每条路径的统计")
    print(f"  3. 详细结果      - 每次运行每条路径的详细结果")
    print(f"  4. 目标路径      - 目标路径定义")
    print(f"{'=' * 70}\n")

    return filename


def main():
    # 目标路径定义（105分支版本）
    target_paths = [
    {9, 10, 11, 13, 14, 15, 16, 18, 19, 31, 32, 33, 34, 36, 75, 78, 81, 83, 84, 85, 86, 87},
    {25, 26, 27, 29, 30, 33, 37, 42, 52, 53, 56, 57, 58, 61, 62, 88, 93, 95, 96, 97},
    {16, 31, 32, 33, 35, 36, 51, 52, 53, 57, 59, 62, 75, 78, 81, 83, 84, 85, 86, 87},
    {2, 5, 6, 7, 9, 10, 31, 32, 33, 34, 35, 39, 44, 47, 75, 81, 83, 84, 85, 86, 87},
    {2, 5, 6, 7, 8, 9, 10, 20, 31, 33, 34, 35, 75, 78, 81, 83, 84, 85, 86, 87, 98},
    {6, 9, 10, 11, 14, 15, 16, 18, 19, 31, 34, 35, 36, 64, 65, 76, 77, 79, 80, 82},
    {1, 2, 5, 6, 7, 8, 9, 10, 20, 31, 32, 33, 34, 35, 36, 70, 72, 93, 94, 98, 99},
    {21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 34, 35, 36, 43, 47, 91, 92},
    {21, 24, 25, 26, 27, 29, 30, 37, 42, 52, 53, 56, 57, 62, 63, 88, 93, 95, 96},
    {3, 31, 32, 33, 34, 35, 36, 39, 40, 41, 44, 45, 47, 88, 89, 90, 95, 96, 97},
    {1, 2, 6, 7, 9, 10, 17, 20, 32, 33, 34, 35, 36, 70, 72, 93, 94, 98, 99},
    {3, 4, 20, 33, 36, 54, 58, 60, 61, 63, 70, 72, 88, 89, 90, 95, 96, 97},
    {6, 9, 10, 20, 31, 32, 34, 35, 69, 71, 74, 77, 79, 80, 82, 94, 98, 99},
    {1, 2, 3, 6, 7, 8, 9, 10, 50, 56, 57, 60, 62, 67, 78, 81, 84, 85, 87},
    {1, 2, 3, 6, 7, 8, 9, 10, 12, 17, 20, 51, 52, 53, 56, 57, 62, 70, 72},
    {21, 24, 25, 26, 27, 29, 30, 31, 37, 39, 42, 44, 48, 57, 88, 95, 96},
    {9, 10, 17, 20, 31, 33, 34, 35, 70, 72, 73, 77, 80, 82, 94, 98, 99},
    {9, 10, 11, 16, 18, 19, 32, 66, 69, 75, 78, 81, 83, 84, 85, 86, 87},
    {1, 2, 3, 6, 7, 9, 10, 11, 13, 14, 15, 16, 18, 19, 32, 55, 70, 72},
    {21, 24, 25, 26, 27, 29, 30, 32, 34, 35, 38, 43, 47, 88, 95, 96},
    {3, 32, 39, 40, 41, 44, 45, 47, 49, 88, 89, 90, 95, 96, 97},
    {3, 31, 32, 34, 37, 42, 46, 88, 90, 95, 96, 97},
    {2, 3, 6, 7, 8, 9, 10, 57, 62, 68, 78, 84}
    ]

    print("=" * 70)
    print("Baseline PSO - 105分支版本")
    print("=" * 70)
    print(f"范围: x(2,100), y(1,105), z(1,110)")
    print(f"路径数: {len(target_paths)}")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("程序完成")


if __name__ == "__main__":
    main()