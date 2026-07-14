import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
from torch.distributions import Normal
import numpy as np
import random
import time
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os
import warnings
warnings.filterwarnings('ignore')

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# ===  ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'ACTION_DIM': 3,
    # : [co2, moisture, temp]
    'MIN_VALUES': np.array([1, 1, 2], dtype=np.float32),
    'MAX_VALUES': np.array([200, 200, 150], dtype=np.float32),
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 3,
    'REPLAY_BATCH_SIZE': 64,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'ACTOR_LR': 3e-4,
    'CRITIC_LR': 3e-4,
    'ALPHA_LR': 3e-4,
    'GAMMA': 0.99,
    'TAU': 0.005,
    'NUM_RUNS': 20,
    'TOP_K_SAMPLES': 20,
    'TARGET_PATHS': [
        {9, 10, 11, 13, 14, 15, 16, 18, 19, 31, 32, 33, 34, 36, 75, 78, 81, 83, 84, 85, 86, 87},
        {25, 26, 27, 29, 30, 33, 37, 42, 52, 53, 56, 57, 58, 61, 62, 88, 93, 95, 96, 97},
        {16, 31, 32, 33, 35, 36, 51, 52, 53, 57, 59, 62, 75, 78, 81, 83, 84, 85, 86, 87},
        {2, 5, 6, 7, 9, 10, 31, 32, 33, 34, 35, 39, 44, 47, 75, 81, 83, 84, 85, 86, 87},
        {2, 5, 6, 7, 8, 9, 10, 20, 31, 33, 34, 35, 75, 78, 81, 83, 84, 85, 86, 87, 98},
        {6, 9, 10, 11, 14, 15, 16, 18, 19, 31, 34, 35, 36, 64, 65, 76, 77, 79, 80, 82},
        {1, 2, 5, 6, 7, 8, 9, 10, 20, 31, 32, 33, 34, 35, 36, 70, 72, 93, 94, 98, 99},
        {21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 34, 35, 36, 43, 47, 91, 92},
        {21, 24, 25, 26, 27, 29, 30, 37, 42, 52, 53, 56, 57, 62, 63, 88, 93, 95, 96},
        {3, 31, 32, 33, 34, 35, 36, 39, 40, 41, 44, 45, 47, 88, 89, 90, 95, 96, 97},
        {1, 2, 6, 7, 9, 10, 17, 20, 32, 33, 34, 35, 36, 70, 72, 93, 94, 98, 99},
        {3, 4, 20, 33, 36, 54, 58, 60, 61, 63, 70, 72, 88, 89, 90, 95, 96, 97},
        {6, 9, 10, 20, 31, 32, 34, 35, 69, 71, 74, 77, 79, 80, 82, 94, 98, 99},
        {1, 2, 3, 6, 7, 8, 9, 10, 50, 56, 57, 60, 62, 67, 78, 81, 84, 85, 87},
        {1, 2, 3, 6, 7, 8, 9, 10, 12, 17, 20, 51, 52, 53, 56, 57, 62, 70, 72},
        {21, 24, 25, 26, 27, 29, 30, 31, 37, 39, 42, 44, 48, 57, 88, 95, 96},
        {9, 10, 17, 20, 31, 33, 34, 35, 70, 72, 73, 77, 80, 82, 94, 98, 99},
        {9, 10, 11, 16, 18, 19, 32, 66, 69, 75, 78, 81, 83, 84, 85, 86, 87},
        {1, 2, 3, 6, 7, 9, 10, 11, 13, 14, 15, 16, 18, 19, 32, 55, 70, 72},
        {21, 24, 25, 26, 27, 29, 30, 32, 34, 35, 38, 43, 47, 88, 95, 96},
        {3, 32, 39, 40, 41, 44, 45, 47, 49, 88, 89, 90, 95, 96, 97},
        {3, 31, 32, 34, 37, 42, 46, 88, 90, 95, 96, 97},
        {2, 3, 6, 7, 8, 9, 10, 57, 62, 68, 78, 84}
    ],
}


# ===  ===
def clip_state(state):
    """"""
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return np.clip(state, min_vals, max_vals)


def normalize_state(state):
    """[-1, 1]"""
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return 2 * (state - min_vals) / (max_vals - min_vals) - 1


def denormalize_state(normalized_state):
    """"""
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return (normalized_state + 1) * (max_vals - min_vals) / 2 + min_vals


def coverage_similarity(triggered, target_path):
    if triggered is None:
        return 0.0
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0

    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)


def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']

    if triggered is not None and target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']

    if triggered is not None and len(triggered) > 0:
        reward += config['TRIGGER_BONUS']

    return reward


def execute_Tr(x, y, z):
    # 初始化分支覆盖数组
    b = [0] * 99  # 根据分支数量调整大小

    if ((x * y) / (z + 1) > 150) != ((x * y) / (z + 1) > 200): b[0] = 1
    if ((x * y) / (z + 1) > 150) != ((x * y) / (z * 2 + 1) > 150): b[1] = 2
    if ((x * y) / (z + 1) > 150) != ((x * x) / (z + 1) > 150): b[2] = 3
    if ((x * y) / (z + 1) > 150) != ((x * 2 * y) / (z + 1) > 150): b[3] = 4
    if ((x * y) / (z + 1) > 150) != ((y * y) / (z + 1) > 150): b[4] = 5
    if ((x * y) / (z + 1) > 150) != ((x * y) / (z + 1) > 500): b[5] = 6
    if ((x * y) / (z + 1) > 150) != ((x * 0.5 * y) / (z + 1) > 150): b[6] = 7
    if ((x * y) / (z + 1) > 150) != ((x * y) / (z + 10) > 150): b[7] = 8
    if ((x * y) / (z + 1) > 150) != ((x * y) / (z * z + 1) > 150): b[8] = 9
    if ((x * y) / (z + 1) > 150) != ((x / y) / (z + 1) > 150): b[9] = 10

    # 验证规则2：相对偏差检测
    if ((y - x) < 0.2 * z) != ((y - x * 2) < 0.2 * z): b[10] = 11
    if ((y - x) < 0.2 * z) != ((y - x) < 0.1 * z): b[11] = 12
    if ((y - x) < 0.2 * z) != ((y - x) < 0.3 * z): b[12] = 13
    if ((y - x) < 0.2 * z) != ((y - x) < 0.5 * z): b[13] = 14
    if ((y - x) < 0.2 * z) != ((y - x) < 0.38 * z): b[14] = 15
    if ((y - x) < 0.2 * z) != ((y - x) < 0.2 * z * x): b[15] = 16
    if ((y - x) < 0.2 * z) != ((y * 1.3 - x) < 0.2 * z): b[16] = 17
    if ((y - x) < 0.2 * z) != ((y - x) < 0.2 * x): b[17] = 18
    if ((y - x) < 0.2 * z) != ((y - x) < 0.2 * y): b[18] = 19
    if ((y - x) < 0.2 * z) != ((y * 2 - x) < 0.2 * z): b[19] = 20

    # 验证规则3：立方根关系验证
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2 + y ** 3) < z ** 2): b[20] = 21
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 2) < z ** 2): b[21] = 22
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 1) < z ** 2): b[22] = 23
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 2.9): b[23] = 24
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 1.8 + y ** 3) < z ** 2): b[24] = 25
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 1 + y ** 3) < z ** 2): b[25] = 26
    if ((x ** 3 + y ** 3) < z ** 2) != ((x * 3 + y ** 3) < z ** 2): b[26] = 27
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * 3) < z ** 2): b[27] = 28
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 3): b[28] = 29
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3.2) < z ** 3): b[29] = 30

    # 验证规则6：整数同余检查
    if (int(x) % 3 == int(y) % 3 == int(z) % 3 == 0) != (
            int(x) % 2 == int(y) % 3 == int(z) % 3 == 0): b[30] = 31
    if (int(x) % 3 == int(y) % 3 == int(z) % 3 == 0) != (
            int(x) % 3 == int(y) % 2 == int(z) % 3 == 0): b[31] = 32
    if (int(x) % 3 == int(y) % 3 == int(z) % 3 == 0) != (
            int(x) % 3 == int(y) % 3 == int(z) % 2 == 0): b[32] = 33
    if (int(x) % 3 == int(y) % 3 == int(z) % 3 == 0) != (
            int(x) % 5 == int(y) % 3 == int(z) % 3 == 0): b[33] = 34
    if (int(x) % 3 == int(y) % 3 == int(z) % 3 == 0) != (
            int(x) % 3 == int(y) % 5 == int(z) % 3 == 0): b[34] = 35
    if (int(x) % 3 == int(y) % 3 == int(z) % 3 == 0) != (
            int(x) % 3 == int(y) % 3 == int(z) % 5 == 0): b[35] = 36

    # 验证规则7：比值范围检查
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (z + 0.1)) > 3 and (y / (z + 0.1)) < 0.3): b[36] = 37
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 1 and (y / (z + 0.1)) < 0.3): b[37] = 38
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (x + 0.1)) < 0.3): b[38] = 39
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (z * 1.2 + 0.1)) < 0.3): b[39] = 40
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.5): b[40] = 41
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (z * 0.1)) < 0.3): b[41] = 42
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y * 0.1)) > 3 and (y / (z + 0.1)) < 0.3): b[42] = 43
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x + (y + 0.1)) > 3 and (y / (x + 0.1)) < 0.3): b[43] = 44
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (z * 1.5 + 0.1)) < 0.3): b[44] = 45
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (z + 0)) < 0.3): b[45] = 46
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 or (y / (z + 0.1)) < 0.3): b[46] = 47
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 2)) > 3 and (y / (z + 0.1)) < 0.3): b[47] = 48
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (z + 5)) < 0.3): b[48] = 49

    # 验证规则8：差值阈值检查
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x * x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8): b[49] = 50
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 20 and abs(y - z) > 20 and abs(x - z) < 8): b[50] = 51
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 30 and abs(y - z) > 20 and abs(x - z) < 8): b[51] = 52
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y - z) > 40 and abs(x - z) < 8): b[52] = 53
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y * z - z) > 20 and abs(x - z) < 8): b[53] = 54
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 12): b[54] = 55
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z * 2) < 8): b[55] = 56
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y - z) > 20 and abs(x * 2 - z) < 8): b[56] = 57
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y * 2 - z) > 20 and abs(x - z) < 8): b[57] = 58
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x * 2 - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8): b[58] = 59
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y - z * z) > 20 and abs(x - z) < 8): b[59] = 60
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y * y - z) > 20 and abs(x - z) < 8): b[60] = 61
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y - z) > 20 and abs(x * x - z) < 8): b[61] = 62
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y * x - z) > 20 and abs(x - z) < 8): b[62] = 63

    # 验证规则9：极值范围检查
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x * 2 > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)): b[63] = 64
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 60 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)): b[64] = 65
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 115 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)): b[65] = 66
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 18) and (y > 85 or y < 2) and (z > 180 or z < 40)): b[66] = 67
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 5) and (y > 85 or y < 2) and (z > 180 or z < 40)): b[67] = 68
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 8) and (y > 130 or y < 2) and (z > 180 or z < 40)): b[68] = 69
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 8) and (y > 85 or y < 2) and (z * z > 180 or z < 40)): b[69] = 70
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 30)): b[70] = 71
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 8) and (y > 85 or y < 2) and (z * 50 > 180 or z < 40)): b[71] = 72
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 60)): b[72] = 73
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 8) and (y > 100 or y < 2) and (z > 180 or z < 40)): b[73] = 74

    # 额外的复杂验证逻辑
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.3 + y ** 0.5 > z and x * y > z ** 1.5): b[74] = 75
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.6 + y ** 0.5 > z and x * y > z ** 1.5): b[75] = 76
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.7 > z and x * y > z ** 1.5): b[76] = 77
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z and x * 0.5 > z ** 1.5): b[77] = 78
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            (x ** 0.5) * 2 + y ** 0.5 > z and x * y > z ** 1.5): b[78] = 79
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y * 0.5 > z and x * y > z ** 1.5): b[79] = 80
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z * 2 and x * y > z ** 1.5): b[80] = 81
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z * 0.5 and x * y > z ** 1.5): b[81] = 82
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z and 0.3 * y > z ** 1.5): b[82] = 83
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z and x * 0.1 > z ** 1.5): b[83] = 84
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z and 0.2 * y > z ** 1.5): b[84] = 85
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z and 0.5 * y > z ** 1.5): b[85] = 86
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z and x * y > z ** 8): b[86] = 87

    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 3 < z ** 2 * 4 and x > y): b[87] = 88
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * 3 and x > y): b[88] = 89
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * 2 and x > y): b[89] = 90
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * 4 and x * x > y): b[90] = 91
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * 4 and x * 2 > y): b[91] = 92
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + 50) ** 2 < z ** 2 * 4 and x > y): b[92] = 93
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 3 * 4 and x > y): b[93] = 94
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 3 < z ** 2 * 4 and x > y): b[94] = 95
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 4 < z ** 2 * 4 and x > y): b[95] = 96
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * 1 and x > y): b[96] = 97
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * x and x > y): b[97] = 98
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * y and x > y): b[98] = 99

    # 返回触发的分支索引集合
    triggered = set()
    for i, val in enumerate(b):
        if val > 0:
            triggered.add(val)
    return triggered


# === SAC Actor ===
class GaussianPolicy(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dim=256):
        super(GaussianPolicy, self).__init__()

        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)

        self.mean = nn.Linear(hidden_dim, action_dim)
        self.log_std = nn.Linear(hidden_dim, action_dim)

        self.action_scale = 8.0
        self.action_bias = 0.0

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))

        mean = self.mean(x)
        log_std = self.log_std(x)
        log_std = torch.clamp(log_std, min=-20, max=2)

        return mean, log_std

    def sample(self, state):
        mean, log_std = self.forward(state)
        std = log_std.exp()
        normal = Normal(mean, std)

        x_t = normal.rsample()
        action = torch.tanh(x_t) * self.action_scale + self.action_bias

        log_prob = normal.log_prob(x_t)
        log_prob -= torch.log(self.action_scale * (1 - torch.tanh(x_t).pow(2)) + 1e-6)
        log_prob = log_prob.sum(1, keepdim=True)

        mean = torch.tanh(mean) * self.action_scale + self.action_bias

        return action, log_prob, mean


# === SAC Critic ===
class QNetwork(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dim=256):
        super(QNetwork, self).__init__()

        self.fc1 = nn.Linear(state_dim + action_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, 1)

        self.fc4 = nn.Linear(state_dim + action_dim, hidden_dim)
        self.fc5 = nn.Linear(hidden_dim, hidden_dim)
        self.fc6 = nn.Linear(hidden_dim, 1)

    def forward(self, state, action):
        xu = torch.cat([state, action], 1)

        x1 = F.relu(self.fc1(xu))
        x1 = F.relu(self.fc2(x1))
        q1 = self.fc3(x1)

        x2 = F.relu(self.fc4(xu))
        x2 = F.relu(self.fc5(x2))
        q2 = self.fc6(x2)

        return q1, q2


# ===  ===
class EnhancedReplayBuffer:
    def __init__(self, capacity=50000):
        self.buffer = deque(maxlen=capacity)
        self.experience_info = deque(maxlen=capacity)

    def push(self, state, action, reward, next_state, done, path_idx, similarity):
        self.buffer.append((state, action, reward, next_state, done))
        self.experience_info.append({'path_idx': path_idx, 'similarity': similarity})

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return None

        batch = random.sample(self.buffer, batch_size)
        state, action, reward, next_state, done = map(np.stack, zip(*batch))

        return (
            torch.FloatTensor(state).to(device),
            torch.FloatTensor(action).to(device),
            torch.FloatTensor(reward).unsqueeze(1).to(device),
            torch.FloatTensor(next_state).to(device),
            torch.FloatTensor(done).unsqueeze(1).to(device)
        )

    def get_top_k_per_path(self, num_paths, k=20):
        path_samples = {i: [] for i in range(num_paths)}

        for idx, info in enumerate(self.experience_info):
            path_idx = info['path_idx']
            similarity = info['similarity']
            path_samples[path_idx].append((idx, similarity, self.buffer[idx]))

        top_k_results = {}
        for path_idx in range(num_paths):
            samples = path_samples[path_idx]
            if len(samples) == 0:
                top_k_results[path_idx] = []
                continue

            samples.sort(key=lambda x: x[1], reverse=True)
            top_k = samples[:k]

            top_k_results[path_idx] = []
            for sample in top_k:
                normalized_state = sample[2][0]
                original_state = denormalize_state(normalized_state)
                original_state_int = np.round(original_state).astype(int)

                # 执行测试覆盖
                co2, moisture, temp = original_state_int
                triggered = execute_Tr(co2, moisture, temp)

                top_k_results[path_idx].append({
                    'state': original_state_int,
                    'similarity': sample[1],
                    'triggered': triggered
                })

        return top_k_results

    def __len__(self):
        return len(self.buffer)


# === SAC ===
class SACAgent:
    def __init__(self, state_dim=3, action_dim=3):
        self.state_dim = state_dim
        self.action_dim = action_dim

        # 初始化网络
        self.policy = GaussianPolicy(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.policy_optimizer = optim.Adam(self.policy.parameters(), lr=EXPERIMENT_CONFIG['ACTOR_LR'])

        self.critic = QNetwork(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_target = QNetwork(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_target.load_state_dict(self.critic.state_dict())
        self.critic_optimizer = optim.Adam(self.critic.parameters(), lr=EXPERIMENT_CONFIG['CRITIC_LR'])

        self.target_entropy = -action_dim
        self.log_alpha = torch.zeros(1, requires_grad=True, device=device)
        self.alpha_optimizer = optim.Adam([self.log_alpha], lr=EXPERIMENT_CONFIG['ALPHA_LR'])

        self.replay_buffer = EnhancedReplayBuffer()
        self.replay_train_count = 0

    def get_action(self, state, deterministic=False):
        normalized_state = normalize_state(state)
        state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)

        with torch.no_grad():
            if deterministic:
                _, _, action = self.policy.sample(state_tensor)
            else:
                action, _, _ = self.policy.sample(state_tensor)

        action = action.cpu().numpy()[0]
        return action

    def store_experience(self, state, action, reward, next_state, done, path_idx, similarity):
        normalized_state = normalize_state(state)
        normalized_next_state = normalize_state(next_state)

        self.replay_buffer.push(
            normalized_state, action, reward,
            normalized_next_state, done, path_idx, similarity
        )

    def replay_train(self):
        batch = self.replay_buffer.sample(EXPERIMENT_CONFIG['REPLAY_BATCH_SIZE'])

        if batch is None:
            return

        state, action, reward, next_state, done = batch

        with torch.no_grad():
            next_action, next_log_prob, _ = self.policy.sample(next_state)

            q1_next, q2_next = self.critic_target(next_state, next_action)
            q_next = torch.min(q1_next, q2_next)
            target_q = reward + (1 - done) * EXPERIMENT_CONFIG['GAMMA'] * (
                    q_next - self.log_alpha.exp() * next_log_prob)

        q1, q2 = self.critic(state, action)
        critic_loss = F.mse_loss(q1, target_q) + F.mse_loss(q2, target_q)

        self.critic_optimizer.zero_grad()
        critic_loss.backward()
        self.critic_optimizer.step()

        new_action, log_prob, _ = self.policy.sample(state)
        q1_new, q2_new = self.critic(state, new_action)
        q_new = torch.min(q1_new, q2_new)

        policy_loss = (self.log_alpha.exp() * log_prob - q_new).mean()

        self.policy_optimizer.zero_grad()
        policy_loss.backward()
        self.policy_optimizer.step()

        alpha_loss = -(self.log_alpha * (log_prob + self.target_entropy).detach()).mean()

        self.alpha_optimizer.zero_grad()
        alpha_loss.backward()
        self.alpha_optimizer.step()

        for param, target_param in zip(self.critic.parameters(), self.critic_target.parameters()):
            target_param.data.copy_(
                EXPERIMENT_CONFIG['TAU'] * param.data + (1 - EXPERIMENT_CONFIG['TAU']) * target_param.data)

        self.replay_train_count += 1

        if self.replay_train_count % 2 == 0:
            alpha_value = self.log_alpha.exp().item()
            print(f"  -> 训练更新 (Run {self.replay_train_count}), Alpha={alpha_value:.4f}")


# === Metric ===
def calculate_run_performance(run_idx, sac_results, training_time, total_steps, update_count, agent):
    """计算单个run的性能指标"""
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    # 收集所有相似度
    all_similarities = []

    # 收集所有奖励
    total_samples = 0
    all_rewards = []
    total_reward = 0

    for path_idx in range(num_paths):
        samples = sac_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']

            total_reward += reward
            all_rewards.append(reward)
            all_similarities.append(similarity)
            total_samples += 1

    # 1. 总奖励
    total_reward = total_reward

    # 2. 平均奖励
    if total_samples > 0:
        average_reward = total_reward / total_samples
    else:
        average_reward = 0

    # 5. 收敛性 (平均相似度)
    if all_similarities:
        convergence = np.mean(all_similarities)
    else:
        convergence = 0

    # 12. 环境适应性 (相似度的标准差倒数)
    if len(all_similarities) > 1:
        environment_adaptability = 1 / (np.std(all_similarities) + 1e-8)
    else:
        environment_adaptability = 0

    # 13. 泛化能力 (平均相似度)
    generalization_ability = convergence

    # 15. 计算效率 (总步数 / 训练时间)
    if training_time > 0:
        computational_efficiency = total_steps / training_time
    else:
        computational_efficiency = 0

    # 16. 策略更新频率
    if training_time > 0:
        policy_update_frequency = update_count / training_time
    else:
        policy_update_frequency = 0

    # 相似度统计
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0

    return {
        'Run': run_idx + 1,

        # 性能指标
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),

        # 相似度统计
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }


# === Excel导出 ===
def export_to_excel(all_sac_results, all_performance_data, target_paths, output_path="SAC_20_run.xlsx"):
    """导出 20 次 run 的 SAC 测试结果到 Excel"""
    print("\n正在导出数据到 Excel...")

    # 初始化收集所有 run 的数据
    all_sac_summary_data = []
    all_sac_detailed_data = []

    # 遍历每次 run 的结果
    for run_idx, (sac_results, performance_data) in enumerate(zip(all_sac_results, all_performance_data)):
        # ===== Sheet1: SACPath 统计 =====
        sac_summary_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = sac_results[path_idx]

            if len(samples) == 0:
                sac_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue

            # 正常结果计算
            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = 'Yes' if perfect_count > 0 else 'No'

            sac_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })

        all_sac_summary_data.extend(sac_summary_data)

        # ===== Sheet2: SACDetailed Sample Data =====
        sac_detailed_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = sac_results[path_idx]

            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']

                sac_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Sample ID': sample_idx + 1,
                    'CO2': int(state[0]),
                    'Moisture': int(state[1]),
                    'Temperature': int(state[2]),
                    'Similarity': round(similarity, 4),
                    'Perfect Coverage': 'Yes' if abs(similarity - 1.0) < 0.001 else 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path))),
                    'Triggered Rules': ', '.join(map(str, sorted(triggered))) if triggered else '',
                    'Intersection Count': len(target_path.intersection(triggered)) if triggered else 0,
                    'Target Rule Count': len(target_path)
                })

        all_sac_detailed_data.extend(sac_detailed_data)

    # 转换为 DataFrame
    sac_summary_df = pd.DataFrame(all_sac_summary_data)
    sac_detailed_df = pd.DataFrame(all_sac_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet1: SACPath 统计
        sac_summary_df.to_excel(writer, sheet_name='SACPath', index=False)

        # Sheet2: SACDetailed Sample Data
        sac_detailed_df.to_excel(writer, sheet_name='SACDetailed Sample Data', index=False)

        # Sheet3: Metric - 性能指标汇总
        selected_columns = [
            'Run',
            'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df_selected = performance_df[selected_columns]
        performance_df_selected.to_excel(writer, sheet_name='Metric', index=False)

        # 获取 workbook 对象进行格式设置
        workbook = writer.book

        # 设置表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        perfect_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # 浅绿色高亮

        # === 设置 Sheet1 样式 ===
        ws1 = writer.sheets['SACPath']
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 高亮完全覆盖的行
        for row_idx in range(2, ws1.max_row + 1):
            if ws1.cell(row_idx, 9).value == 'Yes':  # 第9列是 Perfect Coverage
                for col_idx in range(1, ws1.max_column + 1):
                    ws1.cell(row_idx, col_idx).fill = perfect_fill

        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 18
        ws1.column_dimensions['D'].width = 15
        ws1.column_dimensions['E'].width = 18
        ws1.column_dimensions['F'].width = 15
        ws1.column_dimensions['G'].width = 15
        ws1.column_dimensions['H'].width = 15
        ws1.column_dimensions['I'].width = 18
        ws1.column_dimensions['J'].width = 50

        # === 设置 Sheet2 样式 ===
        ws2 = writer.sheets['SACDetailed Sample Data']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 18
        ws2.column_dimensions['I'].width = 40
        ws2.column_dimensions['J'].width = 40
        ws2.column_dimensions['K'].width = 18
        ws2.column_dimensions['L'].width = 18

        # === 设置 Sheet3 样式 ===
        ws3 = writer.sheets['Metric']
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 调整列宽
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        for col in columns:
            ws3.column_dimensions[col].width = 20

    print(f"文件已成功保存到: {output_path}")
    print(f"  - Sheet1: SACPath 统计共计 {len(all_sac_summary_data)} 条记录")
    print(f"  - Sheet2: SACDetailed Sample Data 共计 {len(all_sac_detailed_data)} 条记录")
    print(f"  - Sheet3: Metric 共计 {len(all_performance_data)} 条记录")


# === 训练工作流 ===
def train_sac_workflow():
    print("=" * 80)
    print("开始SAC训练")
    print("Similarity: 触发分支数 / target paths目标分支数")
    print("=" * 80)

    agent = SACAgent()
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n初始化: 为每个Path生成 {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']} 个样本")
    path_samples = {}
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']

    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            # 生成随机初始状态
            state = np.random.uniform(min_vals, max_vals).astype(np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  Path  {path_idx + 1}/{num_paths}:  生成 {len(samples)} 个样本")

    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size

    print(f"\n开始训练: 每批次{batch_size}个样本, 每个样本{EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}步")
    print(f"总批次数: {num_batches} 批次/Path  x {num_paths} Path  = {num_batches * num_paths} 批次")
    print("-" * 80)

    for batch_idx in range(num_batches):
        print(f"\n批次 {batch_idx + 1}/{num_batches}")

        for path_idx in range(num_paths):
            target_path = target_paths[path_idx]
            batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]

            batch_rewards = []
            batch_similarities = []

            for sample_idx, initial_state in enumerate(batch_samples):
                state = initial_state.copy()
                episode_reward = 0
                final_similarity = 0

                for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                    action = agent.get_action(state, deterministic=False)

                    next_state = state + action
                    next_state = clip_state(next_state)

                    # 计算奖励和相似度
                    co2, moisture, temp = next_state
                    triggered = execute_Tr(co2, moisture, temp)
                    reward = unified_reward_function(triggered, target_path)
                    similarity = coverage_similarity(triggered, target_path)

                    done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)

                    agent.store_experience(
                        state, action, reward, next_state, done,
                        path_idx, similarity
                    )

                    state = next_state
                    episode_reward += reward
                    final_similarity = similarity
                    total_steps += 1

                batch_rewards.append(episode_reward)
                batch_similarities.append(final_similarity)

            avg_reward = np.mean(batch_rewards)
            avg_similarity = np.mean(batch_similarities)
            print(f"  Path {path_idx + 1}: 平均奖励={avg_reward:.2f}, 平均相似度={avg_similarity:.4f}")

        print(f"\n  执行训练更新...")
        agent.replay_train()
        print(f"  缓冲区大小: {len(agent.replay_buffer)}")

    training_time = time.time() - start_time

    print("\n" + "=" * 80)
    print(f"SAC训练完成! 总耗时: {training_time:.2f} 秒, 总步数: {total_steps}")
    print(f"缓冲区大小: {len(agent.replay_buffer)}")
    print(f"训练更新次数: {agent.replay_train_count}")
    print("=" * 80)

    print(f"\n获取每个Path相似度最高的{EXPERIMENT_CONFIG['TOP_K_SAMPLES']}个样本...")
    top_k_results = agent.replay_buffer.get_top_k_per_path(num_paths, EXPERIMENT_CONFIG['TOP_K_SAMPLES'])

    return agent, top_k_results, training_time, total_steps, agent.replay_train_count


# === 主函数 ===
def main():
    print("\n" + "=" * 80)
    print("SAC - 20 run实验")
    print("环境: CO2(2-100), moisture(10-105), temperature(1-110)")
    print("评估Metric性能指标")
    print("=" * 80)

    all_sac_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    # 执行20次run
    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'=' * 80}")
        print(f"开始执行第 {run_idx + 1}/{EXPERIMENT_CONFIG['NUM_RUNS']} 次run")
        print(f"{'=' * 80}")

        # 执行SAC训练
        sac_agent, sac_results, training_time, total_steps, update_count = train_sac_workflow()

        # 计算性能指标
        performance_data = calculate_run_performance(
            run_idx, sac_results, training_time, total_steps, update_count, sac_agent
        )

        # 存储结果
        all_sac_results.append(sac_results)
        all_performance_data.append(performance_data)

        print(f"\n第 {run_idx + 1} 次run完成!")
        print(f"  总奖励: {performance_data['Total Reward']}")
        print(f"  平均奖励: {performance_data['Average Reward']}")
        print(f"  平均相似度: {performance_data['Average Similarity']}")

    # 导出所有结果到Excel(包含20次run)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"SAC_20_run_{timestamp}.xlsx"
    export_to_excel(all_sac_results, all_performance_data, target_paths, output_path)

    # 汇总统计
    print("\n" + "=" * 80)
    print("20次run汇总统计")
    print("=" * 80)

    # 提取所有性能指标
    total_rewards = [p['Total Reward'] for p in all_performance_data]
    average_rewards = [p['Average Reward'] for p in all_performance_data]
    convergences = [p['Convergence'] for p in all_performance_data]
    environment_adaptabilities = [p['Environment Adaptability'] for p in all_performance_data]
    generalization_abilities = [p['Generalization Ability'] for p in all_performance_data]
    computational_efficiencies = [p['Computational Efficiency'] for p in all_performance_data]
    policy_update_frequencies = [p['Policy Update Frequency'] for p in all_performance_data]
    avg_similarities = [p['Average Similarity'] for p in all_performance_data]

    print(f"总奖励统计:")
    print(f"  平均值: {np.mean(total_rewards):.2f}")
    print(f"  标准差: {np.std(total_rewards):.2f}")

    print(f"\n平均奖励统计:")
    print(f"  平均值: {np.mean(average_rewards):.4f}")
    print(f"  标准差: {np.std(average_rewards):.4f}")

    print(f"\n收敛性统计:")
    print(f"  平均值: {np.mean(convergences):.4f}")
    print(f"  标准差: {np.std(convergences):.4f}")

    print(f"\n环境适应性统计:")
    print(f"  平均值: {np.mean(environment_adaptabilities):.4f}")
    print(f"  标准差: {np.std(environment_adaptabilities):.4f}")

    print(f"\n泛化能力统计:")
    print(f"  平均值: {np.mean(generalization_abilities):.4f}")
    print(f"  标准差: {np.std(generalization_abilities):.4f}")

    print(f"\n计算效率统计:")
    print(f"  平均值: {np.mean(computational_efficiencies):.2f}")
    print(f"  标准差: {np.std(computational_efficiencies):.2f}")

    print(f"\n策略更新频率统计:")
    print(f"  平均值: {np.mean(policy_update_frequencies):.4f}")
    print(f"  标准差: {np.std(policy_update_frequencies):.4f}")

    print(f"\n平均相似度统计:")
    print(f"  平均值: {np.mean(avg_similarities):.4f}")
    print(f"  标准差: {np.std(avg_similarities):.4f}")

    print("\n" + "=" * 80)
    print(f"全部 {EXPERIMENT_CONFIG['NUM_RUNS']} 次run已完成!")
    print("=" * 80)


if __name__ == "__main__":
    main()