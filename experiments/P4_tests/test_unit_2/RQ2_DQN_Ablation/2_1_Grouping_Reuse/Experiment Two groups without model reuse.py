import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === 全局状态范围配置 ===
MIN_X = 1
MAX_X = 200
MIN_Y = 1
MAX_Y = 200
MIN_Z = 2
MAX_Z = 150


def normalize_state(state):
    """将状态归一化到[0,1]"""
    x, y, z = state
    normalized_x = (x - MIN_X) / (MAX_X - MIN_X)
    normalized_y = (y - MIN_Y) / (MAX_Y - MIN_Y)
    normalized_z = (z - MIN_Z) / (MAX_Z - MIN_Z)
    return [normalized_x, normalized_y, normalized_z]


def denormalize_state(normalized_state):
    """将归一化状态还原为原始状态"""
    norm_x, norm_y, norm_z = normalized_state
    x = int(norm_x * (MAX_X - MIN_X) + MIN_X)
    y = int(norm_y * (MAX_Y - MIN_Y) + MIN_Y)
    z = int(norm_z * (MAX_Z - MIN_Z) + MIN_Z)
    return [x, y, z]


def jaccard_similarity(set1, set2):
    """计算Jaccard相似度"""
    if not set1 or not set2:
        return 0.0
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0


# === reward function ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward


def execute_Tr(x, y, z):
    # 初始化分支覆盖数组
    b = [0] * 99  # 根据分支数量调整大小

    if ((x * y) / (z + 1) > 150) != ((x * y) / (z + 1) > 200): b[0] = 1
    if ((x * y) / (z + 1) > 150) != ((x * y) / (z * 2 + 1) > 150): b[1] = 2
    if ((x * y) / (z + 1) > 150) != ((x * x) / (z + 1) > 150): b[2] = 3
    if ((x * y) / (z + 1) > 150) != ((x * 2 * y) / (z + 1) > 150): b[3] = 4
    if ((x * y) / (z + 1) > 150) != ((y * y) / (z + 1) > 150): b[4] = 5
    if ((x * y) / (z + 1) > 150) != ((x * y) / (z + 1) > 500): b[5] = 6
    if ((x * y) / (z + 1) > 150) != ((x * 0.5 * y) / (z + 1) > 150): b[6] = 7
    if ((x * y) / (z + 1) > 150) != ((x * y) / (z + 10) > 150): b[7] = 8
    if ((x * y) / (z + 1) > 150) != ((x * y) / (z * z + 1) > 150): b[8] = 9
    if ((x * y) / (z + 1) > 150) != ((x / y) / (z + 1) > 150): b[9] = 10

    # 验证规则2：相对偏差检测
    if ((y - x) < 0.2 * z) != ((y - x * 2) < 0.2 * z): b[10] = 11
    if ((y - x) < 0.2 * z) != ((y - x) < 0.1 * z): b[11] = 12
    if ((y - x) < 0.2 * z) != ((y - x) < 0.3 * z): b[12] = 13
    if ((y - x) < 0.2 * z) != ((y - x) < 0.5 * z): b[13] = 14
    if ((y - x) < 0.2 * z) != ((y - x) < 0.38 * z): b[14] = 15
    if ((y - x) < 0.2 * z) != ((y - x) < 0.2 * z * x): b[15] = 16
    if ((y - x) < 0.2 * z) != ((y * 1.3 - x) < 0.2 * z): b[16] = 17
    if ((y - x) < 0.2 * z) != ((y - x) < 0.2 * x): b[17] = 18
    if ((y - x) < 0.2 * z) != ((y - x) < 0.2 * y): b[18] = 19
    if ((y - x) < 0.2 * z) != ((y * 2 - x) < 0.2 * z): b[19] = 20

    # 验证规则3：立方根关系验证
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2 + y ** 3) < z ** 2): b[20] = 21
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 2) < z ** 2): b[21] = 22
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 1) < z ** 2): b[22] = 23
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 2.9): b[23] = 24
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 1.8 + y ** 3) < z ** 2): b[24] = 25
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 1 + y ** 3) < z ** 2): b[25] = 26
    if ((x ** 3 + y ** 3) < z ** 2) != ((x * 3 + y ** 3) < z ** 2): b[26] = 27
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * 3) < z ** 2): b[27] = 28
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 3): b[28] = 29
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3.2) < z ** 3): b[29] = 30

    # 验证规则6：整数同余检查
    if (int(x) % 3 == int(y) % 3 == int(z) % 3 == 0) != (
            int(x) % 2 == int(y) % 3 == int(z) % 3 == 0): b[30] = 31
    if (int(x) % 3 == int(y) % 3 == int(z) % 3 == 0) != (
            int(x) % 3 == int(y) % 2 == int(z) % 3 == 0): b[31] = 32
    if (int(x) % 3 == int(y) % 3 == int(z) % 3 == 0) != (
            int(x) % 3 == int(y) % 3 == int(z) % 2 == 0): b[32] = 33
    if (int(x) % 3 == int(y) % 3 == int(z) % 3 == 0) != (
            int(x) % 5 == int(y) % 3 == int(z) % 3 == 0): b[33] = 34
    if (int(x) % 3 == int(y) % 3 == int(z) % 3 == 0) != (
            int(x) % 3 == int(y) % 5 == int(z) % 3 == 0): b[34] = 35
    if (int(x) % 3 == int(y) % 3 == int(z) % 3 == 0) != (
            int(x) % 3 == int(y) % 3 == int(z) % 5 == 0): b[35] = 36

    # 验证规则7：比值范围检查
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (z + 0.1)) > 3 and (y / (z + 0.1)) < 0.3): b[36] = 37
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 1 and (y / (z + 0.1)) < 0.3): b[37] = 38
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (x + 0.1)) < 0.3): b[38] = 39
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (z * 1.2 + 0.1)) < 0.3): b[39] = 40
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.5): b[40] = 41
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (z * 0.1)) < 0.3): b[41] = 42
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y * 0.1)) > 3 and (y / (z + 0.1)) < 0.3): b[42] = 43
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x + (y + 0.1)) > 3 and (y / (x + 0.1)) < 0.3): b[43] = 44
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (z * 1.5 + 0.1)) < 0.3): b[44] = 45
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (z + 0)) < 0.3): b[45] = 46
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 or (y / (z + 0.1)) < 0.3): b[46] = 47
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 2)) > 3 and (y / (z + 0.1)) < 0.3): b[47] = 48
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (z + 5)) < 0.3): b[48] = 49

    # 验证规则8：差值阈值检查
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x * x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8): b[49] = 50
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 20 and abs(y - z) > 20 and abs(x - z) < 8): b[50] = 51
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 30 and abs(y - z) > 20 and abs(x - z) < 8): b[51] = 52
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y - z) > 40 and abs(x - z) < 8): b[52] = 53
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y * z - z) > 20 and abs(x - z) < 8): b[53] = 54
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 12): b[54] = 55
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z * 2) < 8): b[55] = 56
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y - z) > 20 and abs(x * 2 - z) < 8): b[56] = 57
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y * 2 - z) > 20 and abs(x - z) < 8): b[57] = 58
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x * 2 - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8): b[58] = 59
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y - z * z) > 20 and abs(x - z) < 8): b[59] = 60
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y * y - z) > 20 and abs(x - z) < 8): b[60] = 61
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y - z) > 20 and abs(x * x - z) < 8): b[61] = 62
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y * x - z) > 20 and abs(x - z) < 8): b[62] = 63

    # 验证规则9：极值范围检查
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x * 2 > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)): b[63] = 64
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 60 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)): b[64] = 65
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 115 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)): b[65] = 66
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 18) and (y > 85 or y < 2) and (z > 180 or z < 40)): b[66] = 67
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 5) and (y > 85 or y < 2) and (z > 180 or z < 40)): b[67] = 68
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 8) and (y > 130 or y < 2) and (z > 180 or z < 40)): b[68] = 69
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 8) and (y > 85 or y < 2) and (z * z > 180 or z < 40)): b[69] = 70
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 30)): b[70] = 71
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 8) and (y > 85 or y < 2) and (z * 50 > 180 or z < 40)): b[71] = 72
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 60)): b[72] = 73
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 8) and (y > 100 or y < 2) and (z > 180 or z < 40)): b[73] = 74

    # 额外的复杂验证逻辑
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.3 + y ** 0.5 > z and x * y > z ** 1.5): b[74] = 75
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.6 + y ** 0.5 > z and x * y > z ** 1.5): b[75] = 76
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.7 > z and x * y > z ** 1.5): b[76] = 77
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z and x * 0.5 > z ** 1.5): b[77] = 78
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            (x ** 0.5) * 2 + y ** 0.5 > z and x * y > z ** 1.5): b[78] = 79
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y * 0.5 > z and x * y > z ** 1.5): b[79] = 80
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z * 2 and x * y > z ** 1.5): b[80] = 81
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z * 0.5 and x * y > z ** 1.5): b[81] = 82
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z and 0.3 * y > z ** 1.5): b[82] = 83
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z and x * 0.1 > z ** 1.5): b[83] = 84
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z and 0.2 * y > z ** 1.5): b[84] = 85
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z and 0.5 * y > z ** 1.5): b[85] = 86
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z and x * y > z ** 8): b[86] = 87

    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 3 < z ** 2 * 4 and x > y): b[87] = 88
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * 3 and x > y): b[88] = 89
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * 2 and x > y): b[89] = 90
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * 4 and x * x > y): b[90] = 91
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * 4 and x * 2 > y): b[91] = 92
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + 50) ** 2 < z ** 2 * 4 and x > y): b[92] = 93
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 3 * 4 and x > y): b[93] = 94
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 3 < z ** 2 * 4 and x > y): b[94] = 95
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 4 < z ** 2 * 4 and x > y): b[95] = 96
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * 1 and x > y): b[96] = 97
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * x and x > y): b[97] = 98
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * y and x > y): b[98] = 99

    # 返回触发的分支索引集合
    triggered = set()
    for i, val in enumerate(b):
        if val > 0:
            triggered.add(val)
    return triggered


# === target path definitions ===
targetPaths = [
    {9, 10, 11, 13, 14, 15, 16, 18, 19, 31, 32, 33, 34, 36, 75, 78, 81, 83, 84, 85, 86, 87},
    {25, 26, 27, 29, 30, 33, 37, 42, 52, 53, 56, 57, 58, 61, 62, 88, 93, 95, 96, 97},
    {16, 31, 32, 33, 35, 36, 51, 52, 53, 57, 59, 62, 75, 78, 81, 83, 84, 85, 86, 87},
    {2, 5, 6, 7, 9, 10, 31, 32, 33, 34, 35, 39, 44, 47, 75, 81, 83, 84, 85, 86, 87},
    {2, 5, 6, 7, 8, 9, 10, 20, 31, 33, 34, 35, 75, 78, 81, 83, 84, 85, 86, 87, 98},
    {6, 9, 10, 11, 14, 15, 16, 18, 19, 31, 34, 35, 36, 64, 65, 76, 77, 79, 80, 82},
    {1, 2, 5, 6, 7, 8, 9, 10, 20, 31, 32, 33, 34, 35, 36, 70, 72, 93, 94, 98, 99},
    {21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 34, 35, 36, 43, 47, 91, 92},
    {21, 24, 25, 26, 27, 29, 30, 37, 42, 52, 53, 56, 57, 62, 63, 88, 93, 95, 96},
    {3, 31, 32, 33, 34, 35, 36, 39, 40, 41, 44, 45, 47, 88, 89, 90, 95, 96, 97},
    {1, 2, 6, 7, 9, 10, 17, 20, 32, 33, 34, 35, 36, 70, 72, 93, 94, 98, 99},
    {3, 4, 20, 33, 36, 54, 58, 60, 61, 63, 70, 72, 88, 89, 90, 95, 96, 97},
    {6, 9, 10, 20, 31, 32, 34, 35, 69, 71, 74, 77, 79, 80, 82, 94, 98, 99},
    {1, 2, 3, 6, 7, 8, 9, 10, 50, 56, 57, 60, 62, 67, 78, 81, 84, 85, 87},
    {1, 2, 3, 6, 7, 8, 9, 10, 12, 17, 20, 51, 52, 53, 56, 57, 62, 70, 72},
    {21, 24, 25, 26, 27, 29, 30, 31, 37, 39, 42, 44, 48, 57, 88, 95, 96},
    {9, 10, 17, 20, 31, 33, 34, 35, 70, 72, 73, 77, 80, 82, 94, 98, 99},
    {9, 10, 11, 16, 18, 19, 32, 66, 69, 75, 78, 81, 83, 84, 85, 86, 87},
    {1, 2, 3, 6, 7, 9, 10, 11, 13, 14, 15, 16, 18, 19, 32, 55, 70, 72},
    {21, 24, 25, 26, 27, 29, 30, 32, 34, 35, 38, 43, 47, 88, 95, 96},
    {3, 32, 39, 40, 41, 44, 45, 47, 49, 88, 89, 90, 95, 96, 97},
    {3, 31, 32, 34, 37, 42, 46, 88, 90, 95, 96, 97},
    {2, 3, 6, 7, 8, 9, 10, 57, 62, 68, 78, 84}
]

# 转换为目标路径列表
target_paths = [set(path) for path in targetPaths]


# === Path Similarity ===
def compute_path_similarity_matrix(paths):
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix


def group_paths_by_similarity(paths):
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.mean(avg_sim_scores)

    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]
    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)

    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group


# === 鲁棒性计算 ===
def compute_robustness(state, path):
    base = execute_Tr(state[0], state[1], state[2])
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = [state[0] + dx, state[1] + dy, state[2] + dz]
                # 边界裁剪
                neighbor[0] = max(MIN_X, min(MAX_X, neighbor[0]))
                neighbor[1] = max(MIN_Y, min(MAX_Y, neighbor[1]))
                neighbor[2] = max(MIN_Z, min(MAX_Z, neighbor[2]))

                n_trig = execute_Tr(neighbor[0], neighbor[1], neighbor[2])
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def compute_q_value_score(state, similar_model):
    """使用相似模型计算Q值分数: 1-归一化Q值"""
    if similar_model is None:
        return 0.0

    try:
        normalized_state = normalize_state(state)
        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = similar_model(state_tensor)
            max_q_value = torch.max(q_values).item()
            normalized_q = max_q_value / 20.0
            normalized_q = max(0.0, min(1.0, normalized_q))
            return 1.0 - normalized_q
    except:
        return 0.0


# === 样本生成和筛选 ===
def generate_samples_for_similar_paths(similar_group, num_candidates=2000, top_k=200, run_id=1):
    """为相似路径组生成样本（使用3维权重）"""
    SIMILAR_WEIGHTS = [0.55, 0.39, 0.06]

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_similar.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Similar Group Path {path_id} - Run {run_id}\n")
            f.write("weather time_period z\tScore\tSimilarity\tLengthDiff\tRobustness\n")
            for s in samples:
                weather, time_period, z = s['state']
                f.write(
                    f"{weather} {time_period} {z}\t{s['score']:.4f}\t{s['similarity']:.4f}\t{s['length_diff']:.4f}\t{s['robustness']:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_grouped"

    for path_idx in similar_group:
        path = target_paths[path_idx]
        path_id = path_idx + 1
        candidate_samples = []
        attempts = 0

        while len(candidate_samples) < num_candidates and attempts < num_candidates * 10:
            attempts += 1

            weather = np.random.randint(MIN_X, MAX_X + 1)
            time_period = np.random.randint(MIN_Y, MAX_Y + 1)
            z = np.random.randint(MIN_Z, MAX_Z + 1)
            state = (weather, time_period, z)
            triggered = execute_Tr(weather, time_period, z)

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)

            candidate_samples.append({
                'state': state,
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            })

        if candidate_samples:
            for sample in candidate_samples:
                score = (SIMILAR_WEIGHTS[0] * sample['similarity'] +
                         SIMILAR_WEIGHTS[1] * sample['length_diff'] +
                         SIMILAR_WEIGHTS[2] * sample['robustness'])
                sample['score'] = score

            candidate_samples.sort(key=lambda x: x['score'], reverse=True)
            selected_samples = candidate_samples[:top_k]
            save_samples(path_id=path_id, samples=selected_samples, base_dir=base_dir)


def generate_samples_for_isolated_paths(isolated_group, similar_model, num_candidates=2000, top_k=200, run_id=1):
    """为孤立路径组生成样本（使用4维权重，包含Q值）"""
    ISOLATED_WEIGHTS = [0.18, 0.21, 0.32, 0.29]

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_isolated.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Isolated Group Path {path_id} - Run {run_id}\n")
            f.write("weather time_period z\tScore\tSimilarity\tLengthDiff\tRobustness\tQValueScore\n")
            for s in samples:
                weather, time_period, z = s['state']
                f.write(
                    f"{weather} {time_period} {z}\t{s['score']:.4f}\t{s['similarity']:.4f}\t{s['length_diff']:.4f}\t{s['robustness']:.4f}\t{s['q_value_score']:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_grouped"

    for path_idx in isolated_group:
        path = target_paths[path_idx]
        path_id = path_idx + 1
        candidate_samples = []
        attempts = 0

        while len(candidate_samples) < num_candidates and attempts < num_candidates * 10:
            attempts += 1

            weather = np.random.randint(MIN_X, MAX_X + 1)
            time_period = np.random.randint(MIN_Y, MAX_Y + 1)
            z = np.random.randint(MIN_Z, MAX_Z + 1)
            state = (weather, time_period, z)
            triggered = execute_Tr(weather, time_period, z)

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            q_score = compute_q_value_score(state, similar_model)

            candidate_samples.append({
                'state': state,
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'q_value_score': q_score,
                'triggered': triggered
            })

        if candidate_samples:
            for sample in candidate_samples:
                score = (ISOLATED_WEIGHTS[0] * sample['similarity'] +
                         ISOLATED_WEIGHTS[1] * sample['length_diff'] +
                         ISOLATED_WEIGHTS[2] * sample['robustness'] +
                         ISOLATED_WEIGHTS[3] * sample['q_value_score'])
                sample['score'] = score

            candidate_samples.sort(key=lambda x: x['score'], reverse=True)
            selected_samples = candidate_samples[:top_k]
            save_samples(path_id=path_id, samples=selected_samples, base_dir=base_dir)


# === 分组经验回放(优先级经验回放) ===
class GroupExperienceReplay:
    def __init__(self, capacity=20000):
        self.capacity = capacity
        self.buffer = deque(maxlen=self.capacity)
        self.priorities = deque(maxlen=self.capacity)

    def append(self, experience):
        self.buffer.append(experience)
        self.priorities.append(experience[-1])  # TD-error作为优先级

    def sample(self, batch_size, alpha=0.6):
        """优先级经验回放采样"""
        priorities = np.array(self.priorities) ** alpha
        probabilities = priorities / np.sum(priorities)

        batch_size = min(batch_size, len(self.buffer))
        batch_indices = np.random.choice(len(self.buffer), batch_size, replace=False, p=probabilities)
        batch = [self.buffer[idx] for idx in batch_indices]
        return batch, batch_indices, probabilities[batch_indices]

    def update_priorities(self, batch_indices, td_errors):
        for idx, td_error in zip(batch_indices, td_errors):
            if idx < len(self.priorities):
                self.priorities[idx] = max(td_error, 1e-6)

    def __len__(self):
        return len(self.buffer)

    def get_high_reward_samples(self, target_path, num_samples=20):
        """获取每个路径的高奖励样本（取前20个）"""
        if len(self.buffer) == 0:
            return []

        samples_with_recalculated_scores = []
        seen_states = set()

        for experience in self.buffer:
            state_tensor = experience[0]
            normalized_state = state_tensor.cpu().numpy().flatten()
            state_tuple = tuple(denormalize_state(normalized_state))

            if state_tuple in seen_states:
                continue
            seen_states.add(state_tuple)

            triggered = execute_Tr(state_tuple[0], state_tuple[1], state_tuple[2])
            new_reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_recalculated_scores.append((state_tuple, new_reward, sim, triggered))

        samples_with_recalculated_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_recalculated_scores[:num_samples]


def load_path_data(file_path):
    path_data = []
    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines[2:]:
            parts = line.strip().split("\t")
            state = tuple(map(int, parts[0].split()))
            path_data.append(state)
    return path_data


# === DQN模型定义 ===
class DQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, 128)
        self.fc2 = nn.Linear(128, 64)
        self.fc3 = nn.Linear(64, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)


# === DQN Agent(优先级经验回放) ===
class DQNAgentWithPER:
    def __init__(self, state_dim, action_dim, replay_buffer, gamma=0.99, epsilon=1.0, epsilon_decay=0.995,
                 epsilon_min=0.1, learning_rate=0.001, alpha=0.6, beta=0.4):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer
        self.alpha = alpha
        self.beta = beta

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        delta_values = [1, 2, 3, 5, -1, -2, -3, -5]
        dim = action_idx // 8
        delta_idx = action_idx % 8
        delta = delta_values[delta_idx]
        if dim == 0:
            return (delta, 0, 0)
        elif dim == 1:
            return (0, delta, 0)
        elif dim == 2:
            return (0, 0, delta)

    def act(self, state):
        """根据epsilon-greedy策略选择动作"""
        if random.random() < self.epsilon:
            return random.randrange(self.action_dim)

        normalized_state = normalize_state(state)
        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)
        return torch.argmax(q_values, dim=1).item()

    def store_transition(self, state, action, reward, next_state, done):
        """存储转换并计算TD-error"""
        normalized_state = normalize_state(state)
        normalized_next_state = normalize_state(next_state)

        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        next_state_tensor = torch.tensor(normalized_next_state, dtype=torch.float32).unsqueeze(0).to(device)

        with torch.no_grad():
            q_values = self.model(state_tensor)
            next_q_values = self.target_model(next_state_tensor)
            max_next_q_values = next_q_values.max(1)[0]
            target_q_values = reward + (self.gamma * max_next_q_values * (1 - done))
            td_error = torch.abs(q_values[0][action] - target_q_values).item()

        self.replay_buffer.append((state_tensor, action, reward, next_state_tensor, done, td_error))
        return td_error

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return

        batch, batch_indices, probabilities = self.replay_buffer.sample(batch_size, alpha=self.alpha)
        states, actions, rewards, next_states, dones, _ = zip(*batch)

        weights = (len(self.replay_buffer) * probabilities) ** (-self.beta)
        weights = weights / weights.max()
        weights = torch.tensor(weights, dtype=torch.float32).to(device)

        states = torch.tensor(np.array([s.cpu().numpy().flatten() for s in states]), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array([ns.cpu().numpy().flatten() for ns in next_states]),
                                   dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))

        td_errors = current_q_values - target_q_values
        weighted_loss = (td_errors.pow(2) * weights).mean()

        self.optimizer.zero_grad()
        weighted_loss.backward()
        self.optimizer.step()

        new_priorities = torch.abs(td_errors).detach().cpu().numpy()
        self.replay_buffer.update_priorities(batch_indices, new_priorities)

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())


# === 分组训练 ===
def train_group(group_paths, path_documents, replay_buffer, batch_size=32, group_name=""):
    """对一组路径进行训练 - 参数: 4x50x3x5"""
    state_dim = 3
    action_dim = 24  # 3维 x 8个delta值

    agent = DQNAgentWithPER(state_dim, action_dim, replay_buffer)

    global_steps = 0
    path_rewards = {}

    print(f"Start training {group_name}, Included Paths: {[idx + 1 for idx in group_paths]}")
    start_time = time.time()

    N_SAMPLES = 200
    BATCH_SIZE = 50
    N_BATCHES = 4
    N_STEPS = 3
    N_REPEATS = 5
    TARGET_UPDATE_EVERY_N_BATCHES = 2

    for path_idx in group_paths:
        file_path = os.path.join(path_documents,
                                 f"path{path_idx + 1}_{'similar' if group_name == 'Similar Group' else 'isolated'}.txt")
        if not os.path.exists(file_path):
            print(f"  文件不存在: Path {path_idx + 1}, 跳过训练")
            continue

        path_data = load_path_data(file_path)
        target_path = target_paths[path_idx]

        if path_idx not in path_rewards:
            path_rewards[path_idx] = 0

        print(f"\n  开始训练路径 {path_idx + 1}, 数据量: {len(path_data)}")

        for repeat in range(N_REPEATS):
            print(f"    第 {repeat + 1}/{N_REPEATS} 轮重复训练")

            batch_count = 0

            for batch_idx in range(N_BATCHES):
                batch_start = batch_idx * BATCH_SIZE
                batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)

                if batch_start >= len(path_data):
                    print(f"      批次 {batch_idx + 1}: 数据不足, 跳过")
                    break

                print(f"      批次 {batch_idx + 1}/{N_BATCHES} (样本 {batch_start}-{batch_end})")

                for sample_idx in range(batch_start, batch_end):
                    if sample_idx >= len(path_data):
                        break

                    state = path_data[sample_idx]
                    prev_state = None
                    prev_triggered = None

                    for step in range(N_STEPS):
                        # 找到合法的动作
                        legal_actions = []
                        for a in range(agent.action_dim):
                            dw, dt, dz = agent.decode_action(a)
                            cand_next = (state[0] + dw, state[1] + dt, state[2] + dz)
                            # 检查是否在边界内
                            if (MIN_X <= cand_next[0] <= MAX_X and
                                    MIN_Y <= cand_next[1] <= MAX_Y and
                                    MIN_Z <= cand_next[2] <= MAX_Z):
                                legal_actions.append(a)

                        if not legal_actions:
                            break

                        if random.random() < agent.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            normalized_state = normalize_state(state)
                            state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        dw, dt, dz = agent.decode_action(action)
                        next_state = (state[0] + dw, state[1] + dt, state[2] + dz)

                        triggered = execute_Tr(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered,
                                                prev_triggered, prev_state)

                        done = (step == N_STEPS - 1)

                        agent.store_transition(state, action, reward, next_state, done)

                        prev_state = state
                        prev_triggered = triggered
                        state = next_state

                        path_rewards[path_idx] += reward
                        global_steps += 1

                print(f"        批次 {batch_idx + 1} 完成, 开始训练模型...")
                if len(agent.replay_buffer) >= batch_size:
                    agent.train(batch_size)

                batch_count += 1

                if batch_count % TARGET_UPDATE_EVERY_N_BATCHES == 0:
                    agent.update_target_model()
                    print(f"        目标网络已更新 (批次 {batch_count})")

        print(f"  路径 {path_idx + 1} 训练完成, 累积奖励: {path_rewards[path_idx]:.2f}")

    training_time = time.time() - start_time
    print(f"\n{group_name} 训练完成, 训练时间: {training_time:.2f} 秒")
    print(f"经验池大小: {len(replay_buffer)}")

    return agent, path_rewards, training_time


# === 主流程 ===
def generate_and_train_grouped_paths_staged(path_documents, similar_group, isolated_group, batch_size=32, run_id=1):
    """分阶段执行: 样本生成, 相似组训练, 隔离组样本生成, 隔离组训练"""

    print(f"\n=== 第 {run_id}/20 轮实验 ===")
    similar_group_paths = [idx + 1 for idx in similar_group]
    isolated_group_paths = [idx + 1 for idx in isolated_group]

    print(f"相似路径组: {similar_group_paths}")
    print(f"孤立路径组: {isolated_group_paths}")

    total_start_time = time.time()

    # 阶段1: 生成相似路径组样本
    print(f"\n[阶段1] 生成相似路径组样本...")
    generate_samples_for_similar_paths(similar_group, num_candidates=2000, top_k=200, run_id=run_id)

    # 阶段2: 训练相似路径组
    print(f"\n[阶段2] 训练相似路径组...")
    similar_replay_buffer = GroupExperienceReplay(capacity=20000)
    similar_agent, similar_path_rewards, similar_training_time = train_group(
        similar_group, path_documents, similar_replay_buffer, batch_size=batch_size, group_name="Similar Group"
    )

    # 阶段3: 使用相似组模型生成隔离路径组样本
    print(f"\n[阶段3] 使用相似组模型生成隔离路径组样本...")
    generate_samples_for_isolated_paths(isolated_group, similar_agent.model, num_candidates=2000, top_k=200,
                                        run_id=run_id)

    # 阶段4: 训练隔离路径组
    print(f"\n[阶段4] 训练隔离路径组...")
    isolated_replay_buffer = GroupExperienceReplay(capacity=20000)
    isolated_agent, isolated_path_rewards, isolated_training_time = train_group(
        isolated_group, path_documents, isolated_replay_buffer, batch_size=batch_size, group_name="Isolated Group"
    )

    total_path_rewards = {**similar_path_rewards, **isolated_path_rewards}
    total_cumulative_reward = sum(total_path_rewards.values())
    total_training_time = time.time() - total_start_time

    print(f"\n=== 第 {run_id}/20 轮实验完成, 总训练时间: {total_training_time:.2f} 秒 ===")

    return similar_agent, isolated_agent, similar_replay_buffer, isolated_replay_buffer, total_cumulative_reward, total_path_rewards, total_training_time


# Excel报告生成
def create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir):
    """创建20轮实验的合并Excel报告"""
    os.makedirs(output_dir, exist_ok=True)

    similar_group_paths = [idx + 1 for idx in similar_group]
    isolated_group_paths = [idx + 1 for idx in isolated_group]

    wb = Workbook()

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    header_color = "4472C4"
    similar_group_color = "E2EFDA"
    isolated_group_color = "FCE4D6"
    stats_color = "FFF2CC"

    # === 工作表1: 路径相似度报告 ===
    ws_paths = wb.active
    ws_paths.title = "路径相似度报告"

    path_headers = ['路径ID', '分组类型'] + [f'第{i}轮' for i in range(1, 21)] + ['平均相似度', '最大相似度',
                                                                                    '最小相似度', '标准差']
    for col, header in enumerate(path_headers, 1):
        cell = ws_paths.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_paths.row_dimensions[1].height = 30

    for path_id in range(1, len(target_paths) + 1):
        row = path_id + 1

        if path_id in similar_group_paths:
            group_type = "高相关路径组"
            row_color = similar_group_color
        elif path_id in isolated_group_paths:
            group_type = "低相关路径组"
            row_color = isolated_group_color
        else:
            group_type = "未分组"
            row_color = "FFFFFF"

        cell = ws_paths.cell(row=row, column=1, value=f"路径 {path_id}")
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        cell.border = thin_border

        cell = ws_paths.cell(row=row, column=2, value=group_type)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        cell.border = thin_border

        path_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            sim = run_data['path_similarities'].get(path_id, {}).get('avg_similarity', 0.0)
            path_similarities.append(sim)

            cell = ws_paths.cell(row=row, column=3 + run_idx, value=round(sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        stats_values = [
            round(np.mean(path_similarities), 4),
            round(np.max(path_similarities), 4),
            round(np.min(path_similarities), 4),
            round(np.std(path_similarities), 4)
        ]

        for i, value in enumerate(stats_values):
            cell = ws_paths.cell(row=row, column=23 + i, value=value)
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
            cell.font = Font(bold=True, size=10)
            cell.border = thin_border

    ws_paths.column_dimensions['A'].width = 13
    ws_paths.column_dimensions['B'].width = 16
    for col in range(3, 23):
        ws_paths.column_dimensions[get_column_letter(col)].width = 10
    for col in range(23, 27):
        ws_paths.column_dimensions[get_column_letter(col)].width = 13

    # === 工作表2: 分组比较报告 ===
    ws_groups = wb.create_sheet("分组比较报告")

    group_headers = ['组名', '包含路径'] + [f'第{i}轮' for i in range(1, 21)] + ['平均相似度', '标准差']
    for col, header in enumerate(group_headers, 1):
        cell = ws_groups.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_groups.row_dimensions[1].height = 30

    row = 2

    # 相似组
    cell = ws_groups.cell(row=row, column=1, value="高相关路径组")
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    cell.border = thin_border

    cell = ws_groups.cell(row=row, column=2, value=','.join(map(str, similar_group_paths)))
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    cell.border = thin_border

    group_similarities = []
    for run_idx, run_data in enumerate(all_runs_data):
        group_sim = np.mean(
            [run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0) for p in similar_group_paths])
        group_similarities.append(group_sim)

        cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(group_sim, 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    cell = ws_groups.cell(row=row, column=23, value=round(np.mean(group_similarities), 4))
    cell.number_format = '0.0000'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border

    cell = ws_groups.cell(row=row, column=24, value=round(np.std(group_similarities), 4))
    cell.number_format = '0.0000'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border

    row += 1

    # 孤立组
    if isolated_group_paths:
        cell = ws_groups.cell(row=row, column=1, value="低相关路径组")
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        cell.border = thin_border

        cell = ws_groups.cell(row=row, column=2, value=','.join(map(str, isolated_group_paths)))
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        cell.border = thin_border

        isolated_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            iso_sim = np.mean(
                [run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0) for p in isolated_group_paths])
            isolated_similarities.append(iso_sim)

            cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(iso_sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        cell = ws_groups.cell(row=row, column=23, value=round(np.mean(isolated_similarities), 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border

        cell = ws_groups.cell(row=row, column=24, value=round(np.std(isolated_similarities), 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border

    ws_groups.column_dimensions['A'].width = 16
    ws_groups.column_dimensions['B'].width = 22
    for col in range(3, 23):
        ws_groups.column_dimensions[get_column_letter(col)].width = 10
    ws_groups.column_dimensions[get_column_letter(23)].width = 14
    ws_groups.column_dimensions[get_column_letter(24)].width = 12

    # === 工作表3: 详细样本数据 ===
    ws_samples = wb.create_sheet("详细样本数据")

    sample_headers = ['轮次', '路径ID', '样本ID', 'Weather', 'TimePeriod', 'Z', '相似度', '触发规则集']
    for col, header in enumerate(sample_headers, 1):
        cell = ws_samples.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_samples.row_dimensions[1].height = 30

    sample_row = 2
    for run_idx, run_data in enumerate(all_runs_data, 1):
        for path_id in range(1, len(target_paths) + 1):
            samples = run_data['path_samples'].get(path_id, [])

            if path_id in similar_group_paths:
                path_color = similar_group_color
            elif path_id in isolated_group_paths:
                path_color = isolated_group_color
            else:
                path_color = "FFFFFF"

            for sample_idx, (state_tuple, reward, sim, triggered) in enumerate(samples, 1):
                weather, time_period, z = state_tuple
                triggered_str = ','.join(map(str, sorted(triggered)))

                cell = ws_samples.cell(row=sample_row, column=1, value=f"第{run_idx}轮")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=path_color, end_color=path_color, fill_type="solid")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=2, value=f"路径{path_id}")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=path_color, end_color=path_color, fill_type="solid")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=3, value=sample_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                for col_offset, value in enumerate([weather, time_period, z]):
                    cell = ws_samples.cell(row=sample_row, column=4 + col_offset, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=7, value=round(sim, 4))
                cell.number_format = '0.0000'
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=8, value=f"{{{triggered_str}}}")
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border

                sample_row += 1

    sample_widths = [13, 13, 11, 10, 12, 8, 12, 45]
    for i, width in enumerate(sample_widths, 1):
        ws_samples.column_dimensions[get_column_letter(i)].width = width

    output_path = os.path.join(output_dir, "20轮实验合并报告.xlsx")
    wb.save(output_path)
    print(f"\n合并Excel报告已生成: {output_path}")


def run_20_times_training():
    """运行20次实验（分阶段：先相似组后孤立组）"""
    model_path_base = r"D:\Experiment\CNN\DQNNEW\saved_models_traffic"
    path_documents = r"D:\Experiment\CNN\DQNNEW\path_samples_grouped"
    output_dir = r"D:\Experiment\CNN\ComparisonExperiment2\excel_reports_traffic"

    os.makedirs(model_path_base, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    similar_group, isolated_group = group_paths_by_similarity(target_paths)
    similar_group_display = [idx + 1 for idx in similar_group]
    isolated_group_display = [idx + 1 for idx in isolated_group]

    print("=" * 60)
    print("20轮实验 - 分阶段训练（先相似组后孤立组）")
    print(
        f"状态范围: weather[{MIN_X},{MAX_X}], time_period[{MIN_Y},{MAX_Y}], z[{MIN_Z},{MAX_Z}]")
    print("策略: 相似组生成样本 -> 相似组训练 -> 孤立组生成样本 -> 孤立组训练")
    print("=" * 60)
    print(f"\n自动分组结果:")
    print(f"相似路径组: {similar_group_display}")
    print(f"孤立路径组: {isolated_group_display}")
    print("\n" + "=" * 60)

    all_runs_data = []
    total_start_time = time.time()

    for run_id in range(1, 21):
        print(f"\n{'=' * 60}")
        print(f"开始第 {run_id}/20 轮实验")
        print(f"{'=' * 60}")

        similar_agent, isolated_agent, similar_buffer, isolated_buffer, total_cumulative_reward, path_rewards, training_time = \
            generate_and_train_grouped_paths_staged(path_documents, similar_group, isolated_group, batch_size=32,
                                                    run_id=run_id)

        # 保存模型
        similar_model_path = os.path.join(model_path_base, f"similar_group_model_run_{run_id}.pth")
        isolated_model_path = os.path.join(model_path_base, f"isolated_group_model_run_{run_id}.pth")

        torch.save({
            'model_state_dict': similar_agent.model.state_dict(),
            'optimizer_state_dict': similar_agent.optimizer.state_dict(),
            'epsilon': similar_agent.epsilon,
            'normalization': {
                'x_range': (MIN_X, MAX_X),
                'y_range': (MIN_Y, MAX_Y),
                'z_range': (MIN_Z, MAX_Z)
            },
            'run_id': run_id,
            'group_type': 'similar_group',
            'group_paths': similar_group_display,
            'pool_size': len(similar_buffer),
            'pool_capacity': 20000,
        }, similar_model_path)

        torch.save({
            'model_state_dict': isolated_agent.model.state_dict(),
            'optimizer_state_dict': isolated_agent.optimizer.state_dict(),
            'epsilon': isolated_agent.epsilon,
            'normalization': {
                'x_range': (MIN_X, MAX_X),
                'y_range': (MIN_Y, MAX_Y),
                'z_range': (MIN_Z, MAX_Z)
            },
            'run_id': run_id,
            'group_type': 'isolated_group',
            'group_paths': isolated_group_display,
            'pool_size': len(isolated_buffer),
            'pool_capacity': 20000,
        }, isolated_model_path)

        print(f"[第{run_id}轮] 模型已保存 (相似组和孤立组)")

        # 收集本轮数据
        run_data = {
            'run_id': run_id,
            'training_time': training_time,
            'total_reward': total_cumulative_reward,
            'path_rewards': path_rewards,
            'path_similarities': {},
            'path_samples': {}
        }

        all_similarities = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            path_id = path_idx + 1

            if path_id in similar_group_display:
                buffer = similar_buffer
            elif path_id in isolated_group_display:
                buffer = isolated_buffer
            else:
                continue

            high_reward_samples = buffer.get_high_reward_samples(target_path, num_samples=20)

            if high_reward_samples:
                similarities = [sim for _, _, sim, _ in high_reward_samples]
                run_data['path_similarities'][path_idx + 1] = {
                    'avg_similarity': np.mean(similarities),
                    'max_similarity': np.max(similarities),
                    'min_similarity': np.min(similarities),
                    'sample_count': len(similarities)
                }
                run_data['path_samples'][path_idx + 1] = high_reward_samples
                all_similarities.extend(similarities)
            else:
                run_data['path_similarities'][path_idx + 1] = {
                    'avg_similarity': 0.0,
                    'max_similarity': 0.0,
                    'min_similarity': 0.0,
                    'sample_count': 0
                }
                run_data['path_samples'][path_idx + 1] = []

        if all_similarities:
            run_data['overall_avg_similarity'] = np.mean(all_similarities)
            run_data['max_similarity'] = np.max(all_similarities)
            run_data['min_similarity'] = np.min(all_similarities)
        else:
            run_data['overall_avg_similarity'] = 0.0
            run_data['max_similarity'] = 0.0
            run_data['min_similarity'] = 0.0

        all_runs_data.append(run_data)

        print(f"[第{run_id}轮] 完成! 总体平均相似度: {run_data['overall_avg_similarity']:.4f}")
        print(f"{'=' * 60}\n")

    total_time = time.time() - total_start_time

    print("\n正在生成合并Excel报告...")
    create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir)

    print("\n" + "=" * 60)
    print("20轮实验全部完成! - 分阶段训练策略")
    print("=" * 60)
    print(f"状态范围:")
    print(f"  weather (X): [{MIN_X}, {MAX_X}]")
    print(f"  time_period (Y): [{MIN_Y}, {MAX_Y}]")
    print(f"  z (Z): [{MIN_Z}, {MAX_Z}]")
    print(f"\n总耗时: {total_time:.2f} 秒 ({total_time / 60:.2f} 分钟)")
    print(f"平均每轮耗时: {total_time / 20:.2f} 秒")
    print(f"\n相似度统计:")
    avg_similarities = [r['overall_avg_similarity'] for r in all_runs_data]
    print(f"  总体平均值: {np.mean(avg_similarities):.4f}")
    print(f"  最大值: {np.max(avg_similarities):.4f}")
    print(f"  最小值: {np.min(avg_similarities):.4f}")
    print(f"  标准差: {np.std(avg_similarities):.4f}")
    print(f"\n所有结果已保存到: {output_dir}")
    print("=" * 60)


if __name__ == "__main__":
    run_20_times_training()