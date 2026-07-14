import os
import random
from collections import deque
import numpy as np
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# --- 全局状态范围配置 ---
STATE_MIN_X, STATE_MAX_X = 1, 200
STATE_MIN_Y, STATE_MAX_Y = 1, 200
STATE_MIN_Z, STATE_MAX_Z = 2, 150

def generate_input():
    return [
        random.randint(STATE_MIN_X, STATE_MAX_X),
        random.randint(STATE_MIN_Y, STATE_MAX_Y),
        random.randint(STATE_MIN_Z, STATE_MAX_Z)
    ]


def execute_Tr(x, y, z):
    # 初始化分支覆盖数组
    b = [0] * 99  # 根据分支数量调整大小

    if ((x * y) / (z + 1) > 150) != ((x * y) / (z + 1) > 200): b[0] = 1
    if ((x * y) / (z + 1) > 150) != ((x * y) / (z * 2 + 1) > 150): b[1] = 2
    if ((x * y) / (z + 1) > 150) != ((x * x) / (z + 1) > 150): b[2] = 3
    if ((x * y) / (z + 1) > 150) != ((x * 2 * y) / (z + 1) > 150): b[3] = 4
    if ((x * y) / (z + 1) > 150) != ((y * y) / (z + 1) > 150): b[4] = 5
    if ((x * y) / (z + 1) > 150) != ((x * y) / (z + 1) > 500): b[5] = 6
    if ((x * y) / (z + 1) > 150) != ((x * 0.5 * y) / (z + 1) > 150): b[6] = 7
    if ((x * y) / (z + 1) > 150) != ((x * y) / (z + 10) > 150): b[7] = 8
    if ((x * y) / (z + 1) > 150) != ((x * y) / (z * z + 1) > 150): b[8] = 9
    if ((x * y) / (z + 1) > 150) != ((x / y) / (z + 1) > 150): b[9] = 10

    # 验证规则2：相对偏差检测
    if ((y - x) < 0.2 * z) != ((y - x * 2) < 0.2 * z): b[10] = 11
    if ((y - x) < 0.2 * z) != ((y - x) < 0.1 * z): b[11] = 12
    if ((y - x) < 0.2 * z) != ((y - x) < 0.3 * z): b[12] = 13
    if ((y - x) < 0.2 * z) != ((y - x) < 0.5 * z): b[13] = 14
    if ((y - x) < 0.2 * z) != ((y - x) < 0.38 * z): b[14] = 15
    if ((y - x) < 0.2 * z) != ((y - x) < 0.2 * z * x): b[15] = 16
    if ((y - x) < 0.2 * z) != ((y * 1.3 - x) < 0.2 * z): b[16] = 17
    if ((y - x) < 0.2 * z) != ((y - x) < 0.2 * x): b[17] = 18
    if ((y - x) < 0.2 * z) != ((y - x) < 0.2 * y): b[18] = 19
    if ((y - x) < 0.2 * z) != ((y * 2 - x) < 0.2 * z): b[19] = 20

    # 验证规则3：立方根关系验证
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 2 + y ** 3) < z ** 2): b[20] = 21
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 2) < z ** 2): b[21] = 22
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 1) < z ** 2): b[22] = 23
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 2.9): b[23] = 24
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 1.8 + y ** 3) < z ** 2): b[24] = 25
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 1 + y ** 3) < z ** 2): b[25] = 26
    if ((x ** 3 + y ** 3) < z ** 2) != ((x * 3 + y ** 3) < z ** 2): b[26] = 27
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y * 3) < z ** 2): b[27] = 28
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3) < z ** 3): b[28] = 29
    if ((x ** 3 + y ** 3) < z ** 2) != ((x ** 3 + y ** 3.2) < z ** 3): b[29] = 30

    # 验证规则6：整数同余检查
    if (int(x) % 3 == int(y) % 3 == int(z) % 3 == 0) != (
            int(x) % 2 == int(y) % 3 == int(z) % 3 == 0): b[30] = 31
    if (int(x) % 3 == int(y) % 3 == int(z) % 3 == 0) != (
            int(x) % 3 == int(y) % 2 == int(z) % 3 == 0): b[31] = 32
    if (int(x) % 3 == int(y) % 3 == int(z) % 3 == 0) != (
            int(x) % 3 == int(y) % 3 == int(z) % 2 == 0): b[32] = 33
    if (int(x) % 3 == int(y) % 3 == int(z) % 3 == 0) != (
            int(x) % 5 == int(y) % 3 == int(z) % 3 == 0): b[33] = 34
    if (int(x) % 3 == int(y) % 3 == int(z) % 3 == 0) != (
            int(x) % 3 == int(y) % 5 == int(z) % 3 == 0): b[34] = 35
    if (int(x) % 3 == int(y) % 3 == int(z) % 3 == 0) != (
            int(x) % 3 == int(y) % 3 == int(z) % 5 == 0): b[35] = 36

    # 验证规则7：比值范围检查
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (z + 0.1)) > 3 and (y / (z + 0.1)) < 0.3): b[36] = 37
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 1 and (y / (z + 0.1)) < 0.3): b[37] = 38
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (x + 0.1)) < 0.3): b[38] = 39
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (z * 1.2 + 0.1)) < 0.3): b[39] = 40
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.5): b[40] = 41
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (z * 0.1)) < 0.3): b[41] = 42
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y * 0.1)) > 3 and (y / (z + 0.1)) < 0.3): b[42] = 43
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x + (y + 0.1)) > 3 and (y / (x + 0.1)) < 0.3): b[43] = 44
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (z * 1.5 + 0.1)) < 0.3): b[44] = 45
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (z + 0)) < 0.3): b[45] = 46
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 or (y / (z + 0.1)) < 0.3): b[46] = 47
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 2)) > 3 and (y / (z + 0.1)) < 0.3): b[47] = 48
    if ((x / (y + 0.1)) > 3 and (y / (z + 0.1)) < 0.3) != (
            (x / (y + 0.1)) > 3 and (y / (z + 5)) < 0.3): b[48] = 49

    # 验证规则8：差值阈值检查
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x * x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8): b[49] = 50
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 20 and abs(y - z) > 20 and abs(x - z) < 8): b[50] = 51
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 30 and abs(y - z) > 20 and abs(x - z) < 8): b[51] = 52
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y - z) > 40 and abs(x - z) < 8): b[52] = 53
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y * z - z) > 20 and abs(x - z) < 8): b[53] = 54
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 12): b[54] = 55
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z * 2) < 8): b[55] = 56
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y - z) > 20 and abs(x * 2 - z) < 8): b[56] = 57
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y * 2 - z) > 20 and abs(x - z) < 8): b[57] = 58
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x * 2 - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8): b[58] = 59
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y - z * z) > 20 and abs(x - z) < 8): b[59] = 60
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y * y - z) > 20 and abs(x - z) < 8): b[60] = 61
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y - z) > 20 and abs(x * x - z) < 8): b[61] = 62
    if (abs(x - y) > 15 and abs(y - z) > 20 and abs(x - z) < 8) != (
            abs(x - y) > 15 and abs(y * x - z) > 20 and abs(x - z) < 8): b[62] = 63

    # 验证规则9：极值范围检查
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x * 2 > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)): b[63] = 64
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 60 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)): b[64] = 65
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 115 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)): b[65] = 66
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 18) and (y > 85 or y < 2) and (z > 180 or z < 40)): b[66] = 67
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 5) and (y > 85 or y < 2) and (z > 180 or z < 40)): b[67] = 68
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 8) and (y > 130 or y < 2) and (z > 180 or z < 40)): b[68] = 69
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 8) and (y > 85 or y < 2) and (z * z > 180 or z < 40)): b[69] = 70
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 30)): b[70] = 71
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 8) and (y > 85 or y < 2) and (z * 50 > 180 or z < 40)): b[71] = 72
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 60)): b[72] = 73
    if ((x > 95 or x < 8) and (y > 85 or y < 2) and (z > 180 or z < 40)) != (
            (x > 95 or x < 8) and (y > 100 or y < 2) and (z > 180 or z < 40)): b[73] = 74

    # 额外的复杂验证逻辑
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.3 + y ** 0.5 > z and x * y > z ** 1.5): b[74] = 75
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.6 + y ** 0.5 > z and x * y > z ** 1.5): b[75] = 76
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.7 > z and x * y > z ** 1.5): b[76] = 77
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z and x * 0.5 > z ** 1.5): b[77] = 78
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            (x ** 0.5) * 2 + y ** 0.5 > z and x * y > z ** 1.5): b[78] = 79
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y * 0.5 > z and x * y > z ** 1.5): b[79] = 80
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z * 2 and x * y > z ** 1.5): b[80] = 81
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z * 0.5 and x * y > z ** 1.5): b[81] = 82
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z and 0.3 * y > z ** 1.5): b[82] = 83
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z and x * 0.1 > z ** 1.5): b[83] = 84
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z and 0.2 * y > z ** 1.5): b[84] = 85
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z and 0.5 * y > z ** 1.5): b[85] = 86
    if (x ** 0.5 + y ** 0.5 > z and x * y > z ** 1.5) != (
            x ** 0.5 + y ** 0.5 > z and x * y > z ** 8): b[86] = 87

    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 3 < z ** 2 * 4 and x > y): b[87] = 88
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * 3 and x > y): b[88] = 89
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * 2 and x > y): b[89] = 90
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * 4 and x * x > y): b[90] = 91
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * 4 and x * 2 > y): b[91] = 92
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + 50) ** 2 < z ** 2 * 4 and x > y): b[92] = 93
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 3 * 4 and x > y): b[93] = 94
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 3 < z ** 2 * 4 and x > y): b[94] = 95
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 4 < z ** 2 * 4 and x > y): b[95] = 96
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * 1 and x > y): b[96] = 97
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * x and x > y): b[97] = 98
    if ((x + y) ** 2 < z ** 2 * 4 and x > y) != (
            (x + y) ** 2 < z ** 2 * y and x > y): b[98] = 99

    # 返回触发的分支索引集合
    triggered = set()
    for i, val in enumerate(b):
        if val > 0:
            triggered.add(val)
    return triggered


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0


targetPaths = [
    {9, 10, 11, 13, 14, 15, 16, 18, 19, 31, 32, 33, 34, 36, 75, 78, 81, 83, 84, 85, 86, 87},
    {25, 26, 27, 29, 30, 33, 37, 42, 52, 53, 56, 57, 58, 61, 62, 88, 93, 95, 96, 97},
    {16, 31, 32, 33, 35, 36, 51, 52, 53, 57, 59, 62, 75, 78, 81, 83, 84, 85, 86, 87},
    {2, 5, 6, 7, 9, 10, 31, 32, 33, 34, 35, 39, 44, 47, 75, 81, 83, 84, 85, 86, 87},
    {2, 5, 6, 7, 8, 9, 10, 20, 31, 33, 34, 35, 75, 78, 81, 83, 84, 85, 86, 87, 98},
    {6, 9, 10, 11, 14, 15, 16, 18, 19, 31, 34, 35, 36, 64, 65, 76, 77, 79, 80, 82},
    {1, 2, 5, 6, 7, 8, 9, 10, 20, 31, 32, 33, 34, 35, 36, 70, 72, 93, 94, 98, 99},
    {21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 34, 35, 36, 43, 47, 91, 92},
    {21, 24, 25, 26, 27, 29, 30, 37, 42, 52, 53, 56, 57, 62, 63, 88, 93, 95, 96},
    {3, 31, 32, 33, 34, 35, 36, 39, 40, 41, 44, 45, 47, 88, 89, 90, 95, 96, 97},
    {1, 2, 6, 7, 9, 10, 17, 20, 32, 33, 34, 35, 36, 70, 72, 93, 94, 98, 99},
    {3, 4, 20, 33, 36, 54, 58, 60, 61, 63, 70, 72, 88, 89, 90, 95, 96, 97},
    {6, 9, 10, 20, 31, 32, 34, 35, 69, 71, 74, 77, 79, 80, 82, 94, 98, 99},
    {1, 2, 3, 6, 7, 8, 9, 10, 50, 56, 57, 60, 62, 67, 78, 81, 84, 85, 87},
    {1, 2, 3, 6, 7, 8, 9, 10, 12, 17, 20, 51, 52, 53, 56, 57, 62, 70, 72},
    {21, 24, 25, 26, 27, 29, 30, 31, 37, 39, 42, 44, 48, 57, 88, 95, 96},
    {9, 10, 17, 20, 31, 33, 34, 35, 70, 72, 73, 77, 80, 82, 94, 98, 99},
    {9, 10, 11, 16, 18, 19, 32, 66, 69, 75, 78, 81, 83, 84, 85, 86, 87},
    {1, 2, 3, 6, 7, 9, 10, 11, 13, 14, 15, 16, 18, 19, 32, 55, 70, 72},
    {21, 24, 25, 26, 27, 29, 30, 32, 34, 35, 38, 43, 47, 88, 95, 96},
    {3, 32, 39, 40, 41, 44, 45, 47, 49, 88, 89, 90, 95, 96, 97},
    {3, 31, 32, 34, 37, 42, 46, 88, 90, 95, 96, 97},
    {2, 3, 6, 7, 8, 9, 10, 57, 62, 68, 78, 84}
]


class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 2000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(targetPaths)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }


def compute_robustness(state, path):
    base = execute_Tr(*state)
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                        np.clip(state[0] + dx, STATE_MIN_X, STATE_MAX_X),
                        np.clip(state[1] + dy, STATE_MIN_Y, STATE_MAX_Y),
                        np.clip(state[2] + dz, STATE_MIN_Z, STATE_MAX_Z)
                    ])
                n_trig = execute_Tr(*neighbor)
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = targetPaths[target_path_idx]
    samples = []
    attempts = 0

    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(STATE_MIN_X, STATE_MAX_X),
            random.randint(STATE_MIN_Y, STATE_MAX_Y),
            random.randint(STATE_MIN_Z, STATE_MAX_Z)
        ])
        triggered = execute_Tr(*state)

        if not triggered:
            continue

        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)

        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }

        samples.append(sample_data)

    return samples


def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample[
                'robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)

        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]

    return selected_samples


def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = targetPaths[target_path_idx]
        samples = []
        attempts = 0

        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(STATE_MIN_X, STATE_MAX_X),
                random.randint(STATE_MIN_Y, STATE_MAX_Y),
                random.randint(STATE_MIN_Z, STATE_MAX_Z)
            ])
            triggered = execute_Tr(*state)

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)

            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)

        random.shuffle(samples)
        return samples[:config.top_k_samples]

    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")

        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)


def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}

    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates

    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}

        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config,
                                                         shared_candidates[path_idx])

            strategy_results[path_idx] = samples

        results[strategy_name] = strategy_results

    return results


def analyze_fitness_values(results, config):
    analysis_results = {}

    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []

        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])

        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }

        all_scores = []

        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])

        scores_array = np.array(all_scores)

        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)

        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })

        analysis_results[strategy_name] = analysis

    return analysis_results


def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []

    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)

    df = pd.DataFrame(df_data)
    return df, analysis_results


def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []

    print(f"Starting {num_runs} experiments...")

    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")

        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)

        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)

        print(f"Experiment {run_idx + 1} completed.")

    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df


def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')

    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width

    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []

    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]

        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)

    stats_df = pd.DataFrame(stats_data)

    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Results saved to: {output_path}")


def main():
    results_df = run_multiple_experiments(num_runs=20)

    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")

    save_results_to_excel(results_df, output_path)

    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)

    return results_df, output_path


if __name__ == "__main__":
    results, output_file = main()