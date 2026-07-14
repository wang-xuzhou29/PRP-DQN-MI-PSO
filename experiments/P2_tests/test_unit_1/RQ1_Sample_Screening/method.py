import os
import random
from collections import deque
import numpy as np
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# --- 全局状态范围配置 ---
STATE_MIN_X, STATE_MAX_X = 1, 128
STATE_MIN_Y, STATE_MAX_Y = 1, 128
STATE_MIN_Z, STATE_MAX_Z = 1, 128
def generate_input():
    return [
        random.randint(STATE_MIN_X, STATE_MAX_X),
        random.randint(STATE_MIN_Y, STATE_MAX_Y),
        random.randint(STATE_MIN_Z, STATE_MAX_Z)
    ]


def execute_Tr(a):
    x, y, z = int(a[0]), int(a[1]), int(a[2])
    triggered = set()

    # Rule Group 1: (x > y) related
    if (x > y) != (x > 5):
        triggered.add(1)
    if (x > y) != (x * x > y):
        triggered.add(2)
    if (x > y) != (x > y * y):
        triggered.add(3)

        # Rule Group 2: (x > z) related
    if (x > z) != (x > 10):
        triggered.add(4)
    if (x > z) != (x * x > z):
        triggered.add(5)
    if (x > z) != (x > z * z):
        triggered.add(6)

        # Rule Group 3: (y > z) related
    if (y > z) != (y > 8):
        triggered.add(7)
    if (y > z) != (y * y > z):
        triggered.add(8)
    if (y > z) != (y > z * z):
        triggered.add(9)
    if (y > z) != (10 > z):
        triggered.add(10)

        # Rule Group 4: (x + y <= z) related
    if (x + y <= z) != (x + y <= z * x):
        triggered.add(11)
    if (x + y <= z) != (x + y <= z * y):
        triggered.add(12)
    if (x + y <= z) != (x * y <= z * z):
        triggered.add(13)
    if (x + y <= z) != (x - y <= z):
        triggered.add(14)

        # 修正后的规则 15：安全处理除以零
    cond_xy_le_z = (x + y <= z)
    cond_x_div_y_le_z = False
    if y != 0:
        cond_x_div_y_le_z = (x / y <= z)

    if cond_xy_le_z != cond_x_div_y_le_z:
        triggered.add(15)

    if (x + y <= z) != (x + y <= 15):
        triggered.add(16)
    if (x + y <= z) != (x + y <= 20):
        triggered.add(17)
    if (x + y <= z) != (x + 5 <= z):
        triggered.add(18)
    if (x + y <= z) != (10 + y <= z):
        triggered.add(19)
    if (x + y <= z) != (x + 8 <= z):
        triggered.add(20)

        # Rule Group 5: (x == y == z) related
    if (x == y == z) != (x <= y == z):
        triggered.add(21)
    if (x == y == z) != (x == y != z):
        triggered.add(22)
    if (x == y == z) != (x != y == z):
        triggered.add(23)

    if (x == y == z) != (x == y <= z):
        triggered.add(24)

        # Rule Group 6: Modulo operations
    if ((x % 2 + y % 2 + z % 2) >= 2) != ((x % 3 + y % 2 + z % 2) >= 2):
        triggered.add(25)
    if ((x % 2 + y % 2 + z % 2) >= 2) != ((x % 2 + y % 3 + z % 2) >= 2):
        triggered.add(26)
    if ((x % 2 + y % 2 + z % 2) >= 2) != ((x % 2 + y % 2 + z % 3) >= 2):
        triggered.add(27)
    if ((x % 2 + y % 2 + z % 2) >= 2) != ((x % 2 + y % 2 + z % 2) >= 1):
        triggered.add(28)
    if ((x % 2 + y % 2 + z % 2) >= 2) != ((x % 2 + y % 2 + z % 2) >= 3):
        triggered.add(29)
    if ((x % 2 + y % 2 + z % 2) >= 2) != ((x % 2 + y % 5 + z % 2) >= 2):
        triggered.add(30)
    if ((x % 2 + y % 2 + z % 2) >= 2) != ((x % 5 + y % 2 + z % 2) >= 2):
        triggered.add(31)
    if ((x % 2 + y % 2 + z % 2) >= 2) != ((x % 2 + y % 2 + z % 5) >= 2):
        triggered.add(32)

        # Rule Group 7: Quadratic equation discriminant like conditions
    cond_main_part = (x != 0 and (y * y - 4 * x * z == 0))

    if cond_main_part != (x != 0 and (y * y - 4 * x * z != 0)):
        triggered.add(33)
    if cond_main_part != (x != 0 and (y * y - 4 * x * z >= 0)):
        triggered.add(34)
    if cond_main_part != (x != 0 and (y * y - 4 * x * z <= 0)):
        triggered.add(35)

    # Rule Group 8: System of equations like conditions
    cond_eq_main_part = (x + y == z and y + z == 2 * x)

    if cond_eq_main_part != (x + y != z and y + z == 2 * x):
        triggered.add(36)

    if cond_eq_main_part != (x + y >= z and y + z == 2 * x):
        triggered.add(37)

    if cond_eq_main_part != (x + y == z and y + z != 2 * x):
        triggered.add(38)
    if cond_eq_main_part != (x + y == z or y + z == 2 * x):
        triggered.add(39)

    return triggered


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0


targetPaths = [
    {1, 2, 4, 5, 7, 8, 11, 12, 13, 14, 15, 16, 17, 18, 19, 24, 25, 26, 27, 32, 33, 35},
    {3, 6, 7, 8, 11, 12, 13, 14, 15, 17, 25, 26, 29, 30, 31, 33, 35},
    {1, 2, 6, 9, 10, 11, 12, 14, 15, 25, 26, 27, 30, 31, 33, 34, 36, 37, 39},
    {30, 1, 2, 4, 5, 33, 7, 8, 35, 16, 17, 38, 39, 26, 29},
    {3, 4, 5, 7, 8, 11, 12, 13, 14, 15, 16, 17, 18, 25, 26, 27, 28, 32, 33, 35},
    {1, 2, 4, 5, 9, 10, 11, 12, 13, 14, 15, 16, 18, 25, 26, 27, 28, 30, 32, 33, 34},
    {1, 2, 4, 5, 7, 8, 11, 12, 13, 14, 15, 16, 17, 18, 19, 21, 29, 30, 32, 33, 35},
    {3, 6, 7, 8, 11, 12, 13, 15, 17, 25, 27, 28, 31, 32, 33, 35},
    {3, 4, 5, 7, 8, 11, 12, 13, 14, 15, 16, 17, 18, 27, 28, 30, 31, 33, 35},
    {1, 2, 4, 5, 7, 8, 11, 12, 13, 14, 15, 16, 18, 27, 30, 33, 35},
    {30, 31, 32, 3, 4, 5, 33, 7, 8, 35, 16, 17, 26, 27, 28},
    {1, 2, 4, 5, 9, 10, 11, 12, 13, 14, 15, 16, 18, 25, 27, 28, 30, 31, 33, 35},
    {3, 4, 5, 7, 8, 11, 12, 13, 14, 15, 16, 17, 18, 25, 28, 30, 31, 33, 35},
    {1, 2, 4, 5, 7, 8, 11, 12, 13, 14, 15, 16, 17, 18, 25, 26, 27, 28, 30, 31, 32, 33, 34},
    {30, 31, 32, 3, 6, 7, 8, 33, 35, 11, 12, 14, 15, 27, 28}
]


class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 2000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(targetPaths)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }


def compute_robustness(state, path):
    base = execute_Tr(*state)
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                        np.clip(state[0] + dx, STATE_MIN_X, STATE_MAX_X),
                        np.clip(state[1] + dy, STATE_MIN_Y, STATE_MAX_Y),
                        np.clip(state[2] + dz, STATE_MIN_Z, STATE_MAX_Z)
                    ])
                n_trig = execute_Tr(*neighbor)
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = targetPaths[target_path_idx]
    samples = []
    attempts = 0

    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(STATE_MIN_X, STATE_MAX_X),
            random.randint(STATE_MIN_Y, STATE_MAX_Y),
            random.randint(STATE_MIN_Z, STATE_MAX_Z)
        ])
        triggered = execute_Tr(*state)

        if not triggered:
            continue

        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)

        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }

        samples.append(sample_data)

    return samples


def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample[
                'robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)

        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]

    return selected_samples


def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = targetPaths[target_path_idx]
        samples = []
        attempts = 0

        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(STATE_MIN_X, STATE_MAX_X),
                random.randint(STATE_MIN_Y, STATE_MAX_Y),
                random.randint(STATE_MIN_Z, STATE_MAX_Z)
            ])
            triggered = execute_Tr(*state)

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)

            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)

        random.shuffle(samples)
        return samples[:config.top_k_samples]

    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")

        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)


def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}

    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates

    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}

        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config,
                                                         shared_candidates[path_idx])

            strategy_results[path_idx] = samples

        results[strategy_name] = strategy_results

    return results


def analyze_fitness_values(results, config):
    analysis_results = {}

    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []

        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])

        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }

        all_scores = []

        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])

        scores_array = np.array(all_scores)

        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)

        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })

        analysis_results[strategy_name] = analysis

    return analysis_results


def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []

    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)

    df = pd.DataFrame(df_data)
    return df, analysis_results


def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []

    print(f"Starting {num_runs} experiments...")

    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")

        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)

        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)

        print(f"Experiment {run_idx + 1} completed.")

    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df


def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')

    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width

    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []

    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]

        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)

    stats_df = pd.DataFrame(stats_data)

    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Results saved to: {output_path}")


def main():
    results_df = run_multiple_experiments(num_runs=20)

    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")

    save_results_to_excel(results_df, output_path)

    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)

    return results_df, output_path


if __name__ == "__main__":
    results, output_file = main()