
import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
from torch.distributions import Normal
import numpy as np
import random
import time
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# ===  ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'ACTION_DIM': 3,
    # : [co2, moisture, temp]
    'MIN_VALUES': np.array([800, 10, 1], dtype=np.float32),
    'MAX_VALUES': np.array([1500, 80, 40], dtype=np.float32),
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 3,
    'REPLAY_BATCH_SIZE': 64,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'ACTOR_LR': 3e-4,
    'CRITIC_LR': 3e-4,
    'ALPHA_LR': 3e-4,
    'GAMMA': 0.99,
    'TAU': 0.005,
    'NUM_RUNS': 20,
    'TOP_K_SAMPLES': 20,
    'TARGET_PATHS': [
        {1, 2, 4, 5, 7, 8, 11, 12, 13, 14, 15, 16, 17, 18, 19, 24, 25, 26, 27, 32, 33, 35},
        {3, 6, 7, 8, 11, 12, 13, 14, 15, 17, 25, 26, 29, 30, 31, 33, 35},
        {1, 2, 6, 9, 10, 11, 12, 14, 15, 25, 26, 27, 30, 31, 33, 34, 36, 37, 39},
        {30, 1, 2, 4, 5, 33, 7, 8, 35, 16, 17, 38, 39, 26, 29},
        {3, 4, 5, 7, 8, 11, 12, 13, 14, 15, 16, 17, 18, 25, 26, 27, 28, 32, 33, 35},
        {1, 2, 4, 5, 9, 10, 11, 12, 13, 14, 15, 16, 18, 25, 26, 27, 28, 30, 32, 33, 34},
        {1, 2, 4, 5, 7, 8, 11, 12, 13, 14, 15, 16, 17, 18, 19, 21, 29, 30, 32, 33, 35},
        {3, 6, 7, 8, 11, 12, 13, 15, 17, 25, 27, 28, 31, 32, 33, 35},
        {3, 4, 5, 7, 8, 11, 12, 13, 14, 15, 16, 17, 18, 27, 28, 30, 31, 33, 35},
        {1, 2, 4, 5, 7, 8, 11, 12, 13, 14, 15, 16, 18, 27, 30, 33, 35},
        {30, 31, 32, 3, 4, 5, 33, 7, 8, 35, 16, 17, 26, 27, 28},
        {1, 2, 4, 5, 9, 10, 11, 12, 13, 14, 15, 16, 18, 25, 27, 28, 30, 31, 33, 35},
        {3, 4, 5, 7, 8, 11, 12, 13, 14, 15, 16, 17, 18, 25, 28, 30, 31, 33, 35},
        {1, 2, 4, 5, 7, 8, 11, 12, 13, 14, 15, 16, 17, 18, 25, 26, 27, 28, 30, 31, 32, 33, 34},
        {30, 31, 32, 3, 6, 7, 8, 33, 35, 11, 12, 14, 15, 27, 28},
    ],
}


# ===  ===
def clip_state(state):
    """"""
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return np.clip(state, min_vals, max_vals)


def normalize_state(state):
    """[-1, 1]"""
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return 2 * (state - min_vals) / (max_vals - min_vals) - 1


def denormalize_state(normalized_state):
    """"""
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return (normalized_state + 1) * (max_vals - min_vals) / 2 + min_vals


def coverage_similarity(triggered, target_path):
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0

    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)


def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']

    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']

    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']

    return reward


def safe_divide(a, b):
    """, """
    return a / b if b != 0 else 0


def execute_Tr(x, y, z):
    x, y, z = int(x), int(y), int(z)
    triggered = set()

    # Rule Group 1: (x > y) related
    if (x > y) != (x > 5):
        triggered.add(1)
    if (x > y) != (x * x > y):
        triggered.add(2)
    if (x > y) != (x > y * y):
        triggered.add(3)

        # Rule Group 2: (x > z) related
    if (x > z) != (x > 10):
        triggered.add(4)
    if (x > z) != (x * x > z):
        triggered.add(5)
    if (x > z) != (x > z * z):
        triggered.add(6)

        # Rule Group 3: (y > z) related
    if (y > z) != (y > 8):
        triggered.add(7)
    if (y > z) != (y * y > z):
        triggered.add(8)
    if (y > z) != (y > z * z):
        triggered.add(9)
    if (y > z) != (10 > z):
        triggered.add(10)

        # Rule Group 4: (x + y <= z) related
    if (x + y <= z) != (x + y <= z * x):
        triggered.add(11)
    if (x + y <= z) != (x + y <= z * y):
        triggered.add(12)
    if (x + y <= z) != (x * y <= z * z):
        triggered.add(13)
    if (x + y <= z) != (x - y <= z):
        triggered.add(14)

        # 修正后的规则 15：安全处理除以零
    cond_xy_le_z = (x + y <= z)
    cond_x_div_y_le_z = False
    if y != 0:
        cond_x_div_y_le_z = (x / y <= z)

    if cond_xy_le_z != cond_x_div_y_le_z:
        triggered.add(15)

    if (x + y <= z) != (x + y <= 15):
        triggered.add(16)
    if (x + y <= z) != (x + y <= 20):
        triggered.add(17)
    if (x + y <= z) != (x + 5 <= z):
        triggered.add(18)
    if (x + y <= z) != (10 + y <= z):
        triggered.add(19)
    if (x + y <= z) != (x + 8 <= z):
        triggered.add(20)

        # Rule Group 5: (x == y == z) related
    if (x == y == z) != (x <= y == z):
        triggered.add(21)
    if (x == y == z) != (x == y != z):
        triggered.add(22)
    if (x == y == z) != (x != y == z):
        triggered.add(23)

    if (x == y == z) != (x == y <= z):
        triggered.add(24)

        # Rule Group 6: Modulo operations
    if ((x % 2 + y % 2 + z % 2) >= 2) != ((x % 3 + y % 2 + z % 2) >= 2):
        triggered.add(25)
    if ((x % 2 + y % 2 + z % 2) >= 2) != ((x % 2 + y % 3 + z % 2) >= 2):
        triggered.add(26)
    if ((x % 2 + y % 2 + z % 2) >= 2) != ((x % 2 + y % 2 + z % 3) >= 2):
        triggered.add(27)
    if ((x % 2 + y % 2 + z % 2) >= 2) != ((x % 2 + y % 2 + z % 2) >= 1):
        triggered.add(28)
    if ((x % 2 + y % 2 + z % 2) >= 2) != ((x % 2 + y % 2 + z % 2) >= 3):
        triggered.add(29)
    if ((x % 2 + y % 2 + z % 2) >= 2) != ((x % 2 + y % 5 + z % 2) >= 2):
        triggered.add(30)
    if ((x % 2 + y % 2 + z % 2) >= 2) != ((x % 5 + y % 2 + z % 2) >= 2):
        triggered.add(31)
    if ((x % 2 + y % 2 + z % 2) >= 2) != ((x % 2 + y % 2 + z % 5) >= 2):
        triggered.add(32)

        # Rule Group 7: Quadratic equation discriminant like conditions
    cond_main_part = (x != 0 and (y * y - 4 * x * z == 0))

    if cond_main_part != (x != 0 and (y * y - 4 * x * z != 0)):
        triggered.add(33)
    if cond_main_part != (x != 0 and (y * y - 4 * x * z >= 0)):
        triggered.add(34)
    if cond_main_part != (x != 0 and (y * y - 4 * x * z <= 0)):
        triggered.add(35)

    # Rule Group 8: System of equations like conditions
    cond_eq_main_part = (x + y == z and y + z == 2 * x)

    if cond_eq_main_part != (x + y != z and y + z == 2 * x):
        triggered.add(36)

    if cond_eq_main_part != (x + y >= z and y + z == 2 * x):
        triggered.add(37)

    if cond_eq_main_part != (x + y == z and y + z != 2 * x):
        triggered.add(38)
    if cond_eq_main_part != (x + y == z or y + z == 2 * x):
        triggered.add(39)

    return triggered

# === SAC Actor ===
class GaussianPolicy(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dim=256):
        super(GaussianPolicy, self).__init__()

        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)

        self.mean = nn.Linear(hidden_dim, action_dim)
        self.log_std = nn.Linear(hidden_dim, action_dim)

        self.action_scale = 8.0
        self.action_bias = 0.0

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))

        mean = self.mean(x)
        log_std = self.log_std(x)
        log_std = torch.clamp(log_std, min=-20, max=2)

        return mean, log_std

    def sample(self, state):
        mean, log_std = self.forward(state)
        std = log_std.exp()
        normal = Normal(mean, std)

        x_t = normal.rsample()
        action = torch.tanh(x_t) * self.action_scale + self.action_bias

        log_prob = normal.log_prob(x_t)
        log_prob -= torch.log(self.action_scale * (1 - torch.tanh(x_t).pow(2)) + 1e-6)
        log_prob = log_prob.sum(1, keepdim=True)

        mean = torch.tanh(mean) * self.action_scale + self.action_bias

        return action, log_prob, mean


# === SAC Critic ===
class QNetwork(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dim=256):
        super(QNetwork, self).__init__()

        self.fc1 = nn.Linear(state_dim + action_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, 1)

        self.fc4 = nn.Linear(state_dim + action_dim, hidden_dim)
        self.fc5 = nn.Linear(hidden_dim, hidden_dim)
        self.fc6 = nn.Linear(hidden_dim, 1)

    def forward(self, state, action):
        xu = torch.cat([state, action], 1)

        x1 = F.relu(self.fc1(xu))
        x1 = F.relu(self.fc2(x1))
        q1 = self.fc3(x1)

        x2 = F.relu(self.fc4(xu))
        x2 = F.relu(self.fc5(x2))
        q2 = self.fc6(x2)

        return q1, q2


# ===  ===
class EnhancedReplayBuffer:
    def __init__(self, capacity=50000):
        self.buffer = deque(maxlen=capacity)
        self.experience_info = deque(maxlen=capacity)

    def push(self, state, action, reward, next_state, done, path_idx, similarity):
        self.buffer.append((state, action, reward, next_state, done))
        self.experience_info.append({'path_idx': path_idx, 'similarity': similarity})

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return None

        batch = random.sample(self.buffer, batch_size)
        state, action, reward, next_state, done = map(np.stack, zip(*batch))

        return (
            torch.FloatTensor(state).to(device),
            torch.FloatTensor(action).to(device),
            torch.FloatTensor(reward).unsqueeze(1).to(device),
            torch.FloatTensor(next_state).to(device),
            torch.FloatTensor(done).unsqueeze(1).to(device)
        )

    def get_top_k_per_path(self, num_paths, k=20):
        path_samples = {i: [] for i in range(num_paths)}

        for idx, info in enumerate(self.experience_info):
            path_idx = info['path_idx']
            similarity = info['similarity']
            path_samples[path_idx].append((idx, similarity, self.buffer[idx]))

        top_k_results = {}
        for path_idx in range(num_paths):
            samples = path_samples[path_idx]
            if len(samples) == 0:
                top_k_results[path_idx] = []
                continue

            samples.sort(key=lambda x: x[1], reverse=True)
            top_k = samples[:k]

            top_k_results[path_idx] = []
            for sample in top_k:
                normalized_state = sample[2][0]
                original_state = denormalize_state(normalized_state)
                original_state_int = np.round(original_state).astype(int)

                # :
                co2, moisture, temp = original_state_int
                triggered = execute_Tr(co2, moisture, temp)

                top_k_results[path_idx].append({
                    'state': original_state_int,
                    'similarity': sample[1],
                    'triggered': triggered
                })

        return top_k_results

    def __len__(self):
        return len(self.buffer)


# === SAC ===
class SACAgent:
    def __init__(self, state_dim=3, action_dim=3):
        self.state_dim = state_dim
        self.action_dim = action_dim

        # :
        self.policy = GaussianPolicy(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.policy_optimizer = optim.Adam(self.policy.parameters(), lr=EXPERIMENT_CONFIG['ACTOR_LR'])

        self.critic = QNetwork(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_target = QNetwork(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_target.load_state_dict(self.critic.state_dict())
        self.critic_optimizer = optim.Adam(self.critic.parameters(), lr=EXPERIMENT_CONFIG['CRITIC_LR'])

        self.target_entropy = -action_dim
        self.log_alpha = torch.zeros(1, requires_grad=True, device=device)
        self.alpha_optimizer = optim.Adam([self.log_alpha], lr=EXPERIMENT_CONFIG['ALPHA_LR'])

        self.replay_buffer = EnhancedReplayBuffer()
        self.replay_train_count = 0

    def get_action(self, state, deterministic=False):
        normalized_state = normalize_state(state)
        state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)

        with torch.no_grad():
            if deterministic:
                _, _, action = self.policy.sample(state_tensor)
            else:
                action, _, _ = self.policy.sample(state_tensor)

        action = action.cpu().numpy()[0]
        return action

    def store_experience(self, state, action, reward, next_state, done, path_idx, similarity):
        normalized_state = normalize_state(state)
        normalized_next_state = normalize_state(next_state)

        self.replay_buffer.push(
            normalized_state, action, reward,
            normalized_next_state, done, path_idx, similarity
        )

    def replay_train(self):
        batch = self.replay_buffer.sample(EXPERIMENT_CONFIG['REPLAY_BATCH_SIZE'])

        if batch is None:
            return

        state, action, reward, next_state, done = batch

        with torch.no_grad():
            next_action, next_log_prob, _ = self.policy.sample(next_state)

            q1_next, q2_next = self.critic_target(next_state, next_action)
            q_next = torch.min(q1_next, q2_next)
            target_q = reward + (1 - done) * EXPERIMENT_CONFIG['GAMMA'] * (
                    q_next - self.log_alpha.exp() * next_log_prob)

        q1, q2 = self.critic(state, action)
        critic_loss = F.mse_loss(q1, target_q) + F.mse_loss(q2, target_q)

        self.critic_optimizer.zero_grad()
        critic_loss.backward()
        self.critic_optimizer.step()

        new_action, log_prob, _ = self.policy.sample(state)
        q1_new, q2_new = self.critic(state, new_action)
        q_new = torch.min(q1_new, q2_new)

        policy_loss = (self.log_alpha.exp() * log_prob - q_new).mean()

        self.policy_optimizer.zero_grad()
        policy_loss.backward()
        self.policy_optimizer.step()

        alpha_loss = -(self.log_alpha * (log_prob + self.target_entropy).detach()).mean()

        self.alpha_optimizer.zero_grad()
        alpha_loss.backward()
        self.alpha_optimizer.step()

        for param, target_param in zip(self.critic.parameters(), self.critic_target.parameters()):
            target_param.data.copy_(
                EXPERIMENT_CONFIG['TAU'] * param.data + (1 - EXPERIMENT_CONFIG['TAU']) * target_param.data)

        self.replay_train_count += 1

        if self.replay_train_count % 2 == 0:
            alpha_value = self.log_alpha.exp().item()
            print(f"  ->  (Run {self.replay_train_count}), Alpha={alpha_value:.4f}")


# === Metric ===
def calculate_run_performance(run_idx, sac_results, training_time, total_steps, update_count, agent):
    """ runMetric"""
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    # Similarity
    all_similarities = []

    # Metric
    total_samples = 0
    all_rewards = []
    total_reward = 0

    for path_idx in range(num_paths):
        samples = sac_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']

            total_reward += reward
            all_rewards.append(reward)
            all_similarities.append(similarity)
            total_samples += 1

    # 1.
    total_reward = total_reward

    # 2.
    if total_samples > 0:
        average_reward = total_reward / total_samples
    else:
        average_reward = 0

    # 5. (Average Similarity)
    if all_similarities:
        convergence = np.mean(all_similarities)
    else:
        convergence = 0

    # 12. (Similarity)
    if len(all_similarities) > 1:
        environment_adaptability = 1 / (np.std(all_similarities) + 1e-8)
    else:
        environment_adaptability = 0

    # 13. (Average Similarity)
    generalization_ability = convergence

    # 15. (/ seconds)
    if training_time > 0:
        computational_efficiency = total_steps / training_time
    else:
        computational_efficiency = 0

    # 16.
    if training_time > 0:
        policy_update_frequency = update_count / training_time
    else:
        policy_update_frequency = 0

    # Similarity
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0

    return {
        'Run': run_idx + 1,

        # Metric
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),

        # Similarity
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }


# === Excel ===
# === Excel导出 ===
def export_to_excel(all_sac_results, all_performance_data, target_paths, output_path="SAC_20 run.xlsx"):
    """导出 20 次 run 的 SAC 测试结果到 Excel"""
    print("\n正在导出数据到 Excel...")

    # 初始化收集所有 run 的数据
    all_sac_summary_data = []
    all_sac_detailed_data = []

    # 遍历每次 run 的结果
    for run_idx, (sac_results, performance_data) in enumerate(zip(all_sac_results, all_performance_data)):
        # ===== Sheet1: SACPath 统计 =====
        sac_summary_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = sac_results[path_idx]

            if len(samples) == 0:
                sac_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue  # 结构修复：这里必须有 continue

            # 正常结果计算
            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = 'Yes' if perfect_count > 0 else 'No'

            sac_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })

        all_sac_summary_data.extend(sac_summary_data)

        # ===== Sheet2: SACDetailed Sample Data =====
        sac_detailed_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = sac_results[path_idx]

            # 结构修复：这里必须有 sample 的遍历循环
            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']

                sac_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Sample ID': sample_idx + 1,
                    'CO2': int(state[0]),
                    'Moisture': int(state[1]),
                    'Temperature': int(state[2]),
                    'Similarity': round(similarity, 4),
                    'Perfect Coverage': 'Yes' if abs(similarity - 1.0) < 0.001 else 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path))),
                    'Triggered Rules': ', '.join(map(str, sorted(triggered))),
                    'Intersection Count': len(target_path.intersection(triggered)),
                    'Target Rule Count': len(target_path)
                })

        all_sac_detailed_data.extend(sac_detailed_data)

    # ================= 结构修复：导出过程必须在所有 Run 循环之后 =================
    # 转换为 DataFrame
    sac_summary_df = pd.DataFrame(all_sac_summary_data)
    sac_detailed_df = pd.DataFrame(all_sac_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet1: SACPath 统计
        sac_summary_df.to_excel(writer, sheet_name='SACPath', index=False)

        # Sheet2: SACDetailed Sample Data
        sac_detailed_df.to_excel(writer, sheet_name='SACDetailed Sample Data', index=False)

        # Sheet3: Metric - 性能指标汇总
        selected_columns = [
            'Run',
            'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df_selected = performance_df[selected_columns]
        performance_df_selected.to_excel(writer, sheet_name='Metric', index=False)

        # 获取 workbook 对象进行格式设置
        workbook = writer.book

        # 设置表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        perfect_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # 浅绿色高亮

        # === 设置 Sheet1 样式 ===
        ws1 = writer.sheets['SACPath']
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 高亮完全覆盖的行
        for row_idx in range(2, ws1.max_row + 1):
            if ws1.cell(row_idx, 9).value == 'Yes':  # 第9列是 Perfect Coverage
                for col_idx in range(1, ws1.max_column + 1):
                    ws1.cell(row_idx, col_idx).fill = perfect_fill

        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 18
        ws1.column_dimensions['D'].width = 15
        ws1.column_dimensions['E'].width = 18
        ws1.column_dimensions['F'].width = 15
        ws1.column_dimensions['G'].width = 15
        ws1.column_dimensions['H'].width = 15
        ws1.column_dimensions['I'].width = 18
        ws1.column_dimensions['J'].width = 50

        # === 设置 Sheet2 样式 ===
        ws2 = writer.sheets['SACDetailed Sample Data']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 18
        ws2.column_dimensions['I'].width = 40
        ws2.column_dimensions['J'].width = 40
        ws2.column_dimensions['K'].width = 18
        ws2.column_dimensions['L'].width = 18

        # === 设置 Sheet3 样式 ===
        ws3 = writer.sheets['Metric']
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 调整列宽
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        for col in columns:
            ws3.column_dimensions[col].width = 20

    print(f"文件已成功保存到: {output_path}")
    print(f"  - Sheet1: SACPath 统计共计 {len(all_sac_summary_data)} 条记录")
    print(f"  - Sheet2: SACDetailed Sample Data 共计 {len(all_sac_detailed_data)} 条记录")
    print(f"  - Sheet3: Metric 共计 {len(all_performance_data)} 条记录")

# ===  ===
def train_sac_workflow():
    print("=" * 80)
    print("SAC")
    print("Similarity:  / target paths")
    print("=" * 80)

    agent = SACAgent()
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n: Path {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']}")
    path_samples = {}
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']

    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            #
            state = np.random.uniform(min_vals, max_vals).astype(np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  Path  {path_idx + 1}/{num_paths}:  {len(samples)} ")

    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size

    print(f"\n: {batch_size},{EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(f": {num_batches} /Path  x {num_paths} Path  = {num_batches * num_paths} ")
    print("-" * 80)

    for batch_idx in range(num_batches):
        print(f"\n {batch_idx + 1}/{num_batches}")

        for path_idx in range(num_paths):
            target_path = target_paths[path_idx]
            batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]

            batch_rewards = []
            batch_similarities = []

            for sample_idx, initial_state in enumerate(batch_samples):
                state = initial_state.copy()
                episode_reward = 0
                final_similarity = 0

                for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                    action = agent.get_action(state, deterministic=False)

                    next_state = state + action
                    next_state = clip_state(next_state)

                    # :
                    co2, moisture, temp = next_state
                    triggered = execute_Tr(co2, moisture, temp)
                    reward = unified_reward_function(triggered, target_path)
                    similarity = coverage_similarity(triggered, target_path)

                    done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)

                    agent.store_experience(
                        state, action, reward, next_state, done,
                        path_idx, similarity
                    )

                    state = next_state
                    episode_reward += reward
                    final_similarity = similarity
                    total_steps += 1

                batch_rewards.append(episode_reward)
                batch_similarities.append(final_similarity)

            avg_reward = np.mean(batch_rewards)
            avg_similarity = np.mean(batch_similarities)
            print(f"  Path {path_idx + 1}: ={avg_reward:.2f}, Average Similarity={avg_similarity:.4f}")

        print(f"\n  ...")
        agent.replay_train()
        print(f"  : {len(agent.replay_buffer)}")

    training_time = time.time() - start_time

    print("\n" + "=" * 80)
    print(f"SACcompleted! Total elapsed time: {training_time:.2f} seconds, : {total_steps}")
    print(f": {len(agent.replay_buffer)}")
    print(f": {agent.replay_train_count}")
    print("=" * 80)

    print(f"\nPath SimilarityMaximum{EXPERIMENT_CONFIG['TOP_K_SAMPLES']}...")
    top_k_results = agent.replay_buffer.get_top_k_per_path(num_paths, EXPERIMENT_CONFIG['TOP_K_SAMPLES'])

    return agent, top_k_results, training_time, total_steps, agent.replay_train_count


# ===  ===
def main():
    print("\n" + "=" * 80)
    print("SAC - 20 run")
    print(": CO2(800-1500), moisture(10-80), temperature(1-40)")
    print("Metric")
    print("=" * 80)

    all_sac_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    # 20
    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'=' * 80}")
        print(f"Start run  {run_idx + 1}/{EXPERIMENT_CONFIG['NUM_RUNS']}  run")
        print(f"{'=' * 80}")

        # SAC
        sac_agent, sac_results, training_time, total_steps, update_count = train_sac_workflow()

        # Metric
        performance_data = calculate_run_performance(
            run_idx, sac_results, training_time, total_steps, update_count, sac_agent
        )

        #
        all_sac_results.append(sac_results)
        all_performance_data.append(performance_data)

        print(f"\nRun  {run_idx + 1}  runcompleted!")
        print(f"  Total Reward: {performance_data['Total Reward']}")
        print(f"  Average Reward: {performance_data['Average Reward']}")
        print(f"  Average Similarity: {performance_data['Average Similarity']}")

    # Excel(20 run)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"SAC_20 run_{timestamp}.xlsx"
    export_to_excel(all_sac_results, all_performance_data, target_paths, output_path)

    #
    print("\n" + "=" * 80)
    print("20 run")
    print("=" * 80)

    # Metric
    total_rewards = [p['Total Reward'] for p in all_performance_data]
    average_rewards = [p['Average Reward'] for p in all_performance_data]
    convergences = [p['Convergence'] for p in all_performance_data]
    environment_adaptabilities = [p['Environment Adaptability'] for p in all_performance_data]
    generalization_abilities = [p['Generalization Ability'] for p in all_performance_data]
    computational_efficiencies = [p['Computational Efficiency'] for p in all_performance_data]
    policy_update_frequencies = [p['Policy Update Frequency'] for p in all_performance_data]
    avg_similarities = [p['Average Similarity'] for p in all_performance_data]

    print(f":")
    print(f"  : {np.mean(total_rewards):.2f}")
    print(f"  Standard deviation: {np.std(total_rewards):.2f}")

    print(f"\n:")
    print(f"  : {np.mean(average_rewards):.4f}")
    print(f"  Standard deviation: {np.std(average_rewards):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(convergences):.4f}")
    print(f"  Standard deviation: {np.std(convergences):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(environment_adaptabilities):.4f}")
    print(f"  Standard deviation: {np.std(environment_adaptabilities):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(generalization_abilities):.4f}")
    print(f"  Standard deviation: {np.std(generalization_abilities):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(computational_efficiencies):.2f}")
    print(f"  Standard deviation: {np.std(computational_efficiencies):.2f}")

    print(f"\n:")
    print(f"  : {np.mean(policy_update_frequencies):.4f}")
    print(f"  Standard deviation: {np.std(policy_update_frequencies):.4f}")

    print(f"\nAverage similarity statistics:")
    print(f"  : {np.mean(avg_similarities):.4f}")
    print(f"  Standard deviation: {np.std(avg_similarities):.4f}")

    print("\n" + "=" * 80)
    print(f" {EXPERIMENT_CONFIG['NUM_RUNS']} completed!")
    print("=" * 80)


if __name__ == "__main__":
    main()