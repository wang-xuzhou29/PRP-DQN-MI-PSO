import os
import random
from collections import deque
import numpy as np
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"


STATE_MIN_X, STATE_MAX_X = 1, 100
STATE_MIN_Y, STATE_MAX_Y = 1, 100
STATE_MIN_Z, STATE_MAX_Z = 1, 100
def generate_input():
    return [
        random.randint(STATE_MIN_X, STATE_MAX_X),
        random.randint(STATE_MIN_Y, STATE_MAX_Y),
        random.randint(STATE_MIN_Z, STATE_MAX_Z)
    ]



def execute_Tr(dx: int, dy: int, dz: int):

 
    MAX_GRID_SIZE = 500.0  
    INITIAL_BATTERY = 1000.0 
    BATTERY_PER_STEP = 1.0 
    SAFE_DISTANCE = 5.0 
    CRITICAL_BATTERY_LEVEL = 100.0 
    TARGET_X, TARGET_Y, TARGET_Z = 450.0, 450.0, 200.0 

    MIN_PLANNING_X = 10.0
    MIN_PLANNING_Y = 15.0
    MIN_PLANNING_Z = 8.0
    CRITICAL_X_VELOCITY = 20.0
    CRITICAL_Y_VELOCITY = 25.0
    CRITICAL_Z_VELOCITY = 15.0

    triggered = set()

    # 模拟环境状态变量，以修复原始代码中的语法错误
    # 这些变量在原始代码中未定义，这里随机生成以确保代码可执行
    current_x = random.uniform(0.0, MAX_GRID_SIZE)
    current_y = random.uniform(0.0, MAX_GRID_SIZE)
    current_z = random.uniform(0.0, MAX_GRID_SIZE)

    # 为了模仿原始代码可能试图比较的'当前位置'与'速度'的关系，
    # 针对第10-15分支中的 'self.y' 采用一个模拟值。
    simulated_y = current_y  # 使用 current_y 作为 self.y 的模拟

    # --- 分支 1-4 ---
    if abs(dx) < MIN_PLANNING_X != abs(dy) < MIN_PLANNING_X: triggered.add(1)
    if abs(dx) < MIN_PLANNING_X != abs(dz) < MIN_PLANNING_X: triggered.add(2)
    if abs(dx) < MIN_PLANNING_X != abs(dx) < MIN_PLANNING_Y: triggered.add(3)
    if abs(dx) < MIN_PLANNING_X != abs(dx) < MIN_PLANNING_Z: triggered.add(4)

    # --- 分支 5-9 ---
    if abs(dz) > MIN_PLANNING_Z * 2 != abs(dx) > MIN_PLANNING_Z * 2: triggered.add(5)
    if abs(dz) > MIN_PLANNING_Z * 2 != abs(dy) > MIN_PLANNING_Z * 2: triggered.add(6)
    if abs(dz) > MIN_PLANNING_Z * 2 != abs(dz) > MIN_PLANNING_X * 2: triggered.add(7)
    if abs(dz) > MIN_PLANNING_Z * 2 != abs(dz) > MIN_PLANNING_Y * 2: triggered.add(8)
    if abs(dz) > MIN_PLANNING_Z * 2 != abs(dz) > MIN_PLANNING_Z: triggered.add(9)

    # --- 分支 10-15 --- (使用 simulated_y 替代 self.y)
    if TARGET_Y > simulated_y and dy < 20 != TARGET_Y > simulated_y and dy < 10: triggered.add(10)
    if TARGET_Y > simulated_y and dy < 20 != TARGET_Y > simulated_y and dy < 30: triggered.add(11)
    if TARGET_Y > simulated_y and dy < 20 != TARGET_Y > simulated_y and dy < 40: triggered.add(12)
    if TARGET_Y > simulated_y and dy < 20 != TARGET_Y > simulated_y and dy < 50: triggered.add(13)
    if TARGET_Y > simulated_y and dy < 20 != TARGET_Y > simulated_y and dx < 20: triggered.add(14)
    if TARGET_Y > simulated_y and dy < 20 != TARGET_Y > simulated_y and dz < 20: triggered.add(15)

    # --- 分支 16-21 ---
    if abs(dy) > CRITICAL_X_VELOCITY * 1.5 != abs(dx) > CRITICAL_X_VELOCITY * 1.5: triggered.add(16)
    if abs(dy) > CRITICAL_X_VELOCITY * 1.5 != abs(dz) > CRITICAL_X_VELOCITY * 1.5: triggered.add(17)
    if abs(dy) > CRITICAL_X_VELOCITY * 1.5 != abs(dy) > CRITICAL_X_VELOCITY: triggered.add(18)
    if abs(dy) > CRITICAL_X_VELOCITY * 1.5 != abs(dy) > CRITICAL_X_VELOCITY * 2: triggered.add(19)
    if abs(dy) > CRITICAL_X_VELOCITY * 1.5 != abs(dy) > CRITICAL_Z_VELOCITY * 1.5: triggered.add(20)
    if abs(dy) > CRITICAL_X_VELOCITY * 1.5 != abs(dy) > CRITICAL_Y_VELOCITY * 1.5: triggered.add(21)

    # --- 分支 22-29 --- (使用 current_x, current_y, current_z 替代未定义的变量)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_X < current_z and dz > CRITICAL_Z_VELOCITY: triggered.add(
        22)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Y < current_z and dz > CRITICAL_Z_VELOCITY: triggered.add(
        23)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Z < current_x and dz > CRITICAL_Z_VELOCITY: triggered.add(
        24)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Z < current_y and dz > CRITICAL_Z_VELOCITY: triggered.add(
        25)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Z < current_z and dx > CRITICAL_Z_VELOCITY: triggered.add(
        26)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Z < current_z and dy > CRITICAL_Z_VELOCITY: triggered.add(
        27)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Z < current_z and dz > CRITICAL_X_VELOCITY: triggered.add(
        28)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Z < current_z and dz > CRITICAL_Y_VELOCITY: triggered.add(
        29)
    return triggered


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0


targetPaths = [
    {1, 2, 3, 4, 10, 11, 12, 13, 14, 15, 24, 25, 26, 27, 28, 29},
    {5, 6, 7, 8, 9, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25},
    {5, 6, 7, 8, 9, 17, 18, 19, 20, 21, 24, 25, 26, 27, 28, 29}
]


class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 2000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(targetPaths)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }


def compute_robustness(state, path):
    base = execute_Tr(*state)
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                        np.clip(state[0] + dx, STATE_MIN_X, STATE_MAX_X),
                        np.clip(state[1] + dy, STATE_MIN_Y, STATE_MAX_Y),
                        np.clip(state[2] + dz, STATE_MIN_Z, STATE_MAX_Z)
                    ])
                n_trig = execute_Tr(*neighbor)
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = targetPaths[target_path_idx]
    samples = []
    attempts = 0

    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(STATE_MIN_X, STATE_MAX_X),
            random.randint(STATE_MIN_Y, STATE_MAX_Y),
            random.randint(STATE_MIN_Z, STATE_MAX_Z)
        ])
        triggered = execute_Tr(*state)

        if not triggered:
            continue

        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)

        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }

        samples.append(sample_data)

    return samples


def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample[
                'robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)

        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]

    return selected_samples


def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = targetPaths[target_path_idx]
        samples = []
        attempts = 0

        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(STATE_MIN_X, STATE_MAX_X),
                random.randint(STATE_MIN_Y, STATE_MAX_Y),
                random.randint(STATE_MIN_Z, STATE_MAX_Z)
            ])
            triggered = execute_Tr(*state)

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)

            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)

        random.shuffle(samples)
        return samples[:config.top_k_samples]

    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")

        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)


def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}

    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates

    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}

        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config,
                                                         shared_candidates[path_idx])

            strategy_results[path_idx] = samples

        results[strategy_name] = strategy_results

    return results


def analyze_fitness_values(results, config):
    analysis_results = {}

    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []

        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])

        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }

        all_scores = []

        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])

        scores_array = np.array(all_scores)

        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)

        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })

        analysis_results[strategy_name] = analysis

    return analysis_results


def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []

    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)

    df = pd.DataFrame(df_data)
    return df, analysis_results


def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []

    print(f"Starting {num_runs} experiments...")

    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")

        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)

        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)

        print(f"Experiment {run_idx + 1} completed.")

    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df


def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')

    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width

    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []

    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]

        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)

    stats_df = pd.DataFrame(stats_data)

    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Results saved to: {output_path}")


def main():
    results_df = run_multiple_experiments(num_runs=20)

    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")

    save_results_to_excel(results_df, output_path)

    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)

    return results_df, output_path


if __name__ == "__main__":
    results, output_file = main()
