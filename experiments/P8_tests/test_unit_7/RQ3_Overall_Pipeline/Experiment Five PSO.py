import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

def execute_validation_rules(a):
    """
    替换原有的编排规则函数
    参数a: 包含3个元素的元组或数组，分别对应path_depth, file_count, access_level
    返回: 触发的规则编号集合
    """
    path_depth, file_count, access_level = float(a[0]), int(a[1]), float(a[2])

    # 使用正确的变量名
    x, y, z = path_depth, file_count, access_level

    triggered = set()
    # 创建一个字典来存储b数组的值，用于跟踪哪些规则被触发
    b = {}
    x_score = max(0, 100 - x * 1.67)

    if ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((42 * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):b[0] = 1
    if ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):b[1] = 2
    if ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * x) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):b[2] = 3
    if ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (x + 1) > 50 and x ** 2 + y ** 2 > z ** 2):b[3] = 4
    if ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (y + 1) > 50 and x ** 2 + y ** 2 > z ** 2):b[4] = 5
    if ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (25 + 1) > 50 and x ** 2 + y ** 2 > z ** 2):b[5] = 6
    if ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z * 1) > 50 and x ** 2 + y ** 2 > z ** 2):b[6] = 7
    if ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 10) > 50 and x ** 2 + y ** 2 > z ** 2):b[7] = 8
    if ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 1) != 50 and x ** 2 + y ** 2 > z ** 2):b[8] = 9
    if ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 1) > 33 and x ** 2 + y ** 2 > z ** 2):b[9] = 10
    if ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2): b[
        10] = 11
    if ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 1) > 50 and y ** 2 + y ** 2 > z ** 2):b[11] = 12
    if ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 1) > 50 and x ** 2 + x ** 2 > z ** 2):b[12] = 13
    if ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 1) > 50 and x ** 2.2 + y ** 2 > z ** 2):b[13] = 14
    if ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2.5 > z ** 2):b[14] = 15
    if ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 1) > 50 and x ** 2 + y ** 1.2 > z ** 2):b[15] = 16
    if ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 1) > 50 and x ** 1.5 + y ** 2 > z ** 2):b[16] = 17
    if ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 != z ** 2):b[17] = 18
    if ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 3):b[18] = 19
    if ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 1) > 50 and 25 ** 2 + y ** 2 > z ** 2):b[19] = 20
    if ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 1) > 50 and x ** 2 + 33 ** 2 > z ** 2):b[20] = 21
    if ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > 42 ** 2):b[21] = 22
    if ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 1.6):b[22] = 23


    if (y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2:
        pattern_type = "151A1"
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45): b[23] = 24
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (54 ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45): b[24] = 25
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2.4 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45): b[25] = 26
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 1.6 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45): b[26] = 27
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2 - z ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45): b[27] = 28
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2.8) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45): b[28] = 29
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 1.6) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45): b[29] = 30
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2 - 42 ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45): b[30] = 31
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (34 + 0.1) < -30 or (abs(z - y) * z) / 100 > 45): b[31] = 32
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (x + 0.1) < -30 or (abs(z - y) * z) / 100 > 45): b[32] = 33
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(z - y) * z) / 100 > 45): b[33] = 34
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 2) < -30 or (abs(z - y) * z) / 100 > 45): b[34] = 35
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) != -30 or (abs(z - y) * z) / 100 > 45): b[35] = 36
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -11 or (abs(z - y) * z) / 100 > 45): b[36] = 37
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 and (abs(z - y) * z) / 100 > 45): b[37] = 38
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[38] = 39
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - x) * z) / 100 > 45): b[39] = 40
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(33 - y) * z) / 100 > 45): b[40] = 41
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - 21) * z) / 100 > 45): b[41] = 42
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * 43) / 100 > 45): b[42] = 43
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * x) / 100 > 45): b[43] = 44
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * y) / 100 > 45): b[44] = 45
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 85 > 45): b[45] = 46
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 != 45): b[46] = 47
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 33): b[47] = 48
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 55): b[48] = 49
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 109 > 45): b[49] = 50
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z - 0.1) < -30 or (abs(z - y) * z) / 100 > 45): b[50] = 51
    if (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45:
        pattern_type = "151A2"
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35) != (
            ((23 + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35): b[51] = 52
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35) != (
            ((x + 33) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35): b[52] = 53
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35) != (
            ((y + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35): b[53] = 54
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35) != (
            ((x + x) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35): b[54] = 55
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35) != (
            ((z + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35): b[55] = 56
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35) != (
            ((x + z) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35): b[56] = 57
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35) != (
            ((x + y) / 2.3) ** 2 < z * 20 and (x * y * x) / 1000 < 35): b[57] = 58
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35) != (
            ((x + y) / 2) ** 2.4 < z * 20 and (x * y * x) / 1000 < 35): b[58] = 59
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35) != (
            ((x + y) / 2) ** 2 != z * 20 and (x * y * x) / 1000 < 35): b[59] = 60
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < x * 20 and (x * y * x) / 1000 < 35): b[60] = 61
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < y * 20 and (x * y * x) / 1000 < 35): b[61] = 62
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < 21 * 20 and (x * y * x) / 1000 < 35): b[62] = 63
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 13 and (x * y * x) / 1000 < 35): b[63] = 64
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 or (x * y * x) / 1000 < 35): b[64] = 65
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (24 * y * x) / 1000 < 35): b[65] = 66
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * 33 * x) / 1000 < 35): b[66] = 67
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (z * y * x) / 1000 < 35): b[67] = 68
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * z * x) / 1000 < 35): b[68] = 69
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * y * y) / 1000 < 35): b[69] = 70
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * x * x) / 1000 < 35): b[70] = 71
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1060 < 35): b[71] = 72
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 != 35): b[72] = 73
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 26): b[73] = 74
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35) != (
            ((x + y) / 2) ** 1.5 < z * 20 and (x * y * x) / 1000 < 35): b[74] = 75
    if ((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35:
        pattern_type = "151A3"
    if (x > 75 and y < 40) != (x > 75 or y < 40): b[75] = 76
    if (x > 75 and y < 40) != (x > 75 and y != 40): b[76] = 77
    if (x > 75 and y < 40) != (x > 75 and y < 26): b[77] = 78
    if (x > 75 and y < 40) != (x != 75 and y < 40): b[78] = 79
    if (x > 75 and y < 40) != (x > 83 and y < 40): b[79] = 80
    if x > 75 and y < 40:
        pattern_type = 152
    if (((z > 80 or z < 20) and y < 60) or ((z > 75 or z < 25) and y < 55)) != (
            ((z != 80 or z < 20) and y < 60) or ((z > 75 or z < 25) and y < 55)): b[80] = 81
    if (((z > 80 or z < 20) and y < 60) or ((z > 75 or z < 25) and y < 55)) != (
            ((z > 62 or z < 20) and y < 60) or ((z > 75 or z < 25) and y < 55)): b[81] = 82
    if (((z > 80 or z < 20) and y < 60) or ((z > 75 or z < 25) and y < 55)) != (
            ((z > 80 and z < 20) and y < 60) or ((z > 75 or z < 25) and y < 55)): b[82] = 83
    if (((z > 80 or z < 20) and y < 60) or ((z > 75 or z < 25) and y < 55)) != (
            ((z > 80 or z != 20) and y < 60) or ((z > 75 or z < 25) and y < 55)): b[83] = 84
    if (((z > 80 or z < 20) and y < 60) or ((z > 75 or z < 25) and y < 55)) != (
            ((z > 80 or z < 11) and y < 60) or ((z > 75 or z < 25) and y < 55)): b[84] = 85
    if (((z > 80 or z < 20) and y < 60) or ((z > 75 or z < 25) and y < 55)) != (
            ((z > 80 or z < 20) or y < 60) or ((z > 75 or z < 25) and y < 55)): b[85] = 86
    if (((z > 80 or z < 20) and y < 60) or ((z > 75 or z < 25) and y < 55)) != (
            ((z > 80 or z < 20) and y != 60) or ((z > 75 or z < 25) and y < 55)): b[86] = 87
    if (((z > 80 or z < 20) and y < 60) or ((z > 75 or z < 25) and y < 55)) != (
            ((z > 80 or z < 20) and y < 44) or ((z > 75 or z < 25) and y < 55)): b[87] = 88
    if (((z > 80 or z < 20) and y < 60) or ((z > 75 or z < 25) and y < 55)) != (
            ((z > 80 or z < 20) and y < 60) and ((z > 75 or z < 25) and y < 55)): b[88] = 89
    if (((z > 80 or z < 20) and y < 60) or ((z > 75 or z < 25) and y < 55)) != (
            ((z > 80 or z < 20) and y < 60) or ((z != 75 or z < 25) and y < 55)): b[89] = 90
    if (((z > 80 or z < 20) and y < 60) or ((z > 75 or z < 25) and y < 55)) != (
            ((z > 80 or z < 20) and y < 60) or ((z > 65 or z < 25) and y < 55)): b[90] = 91
    if (((z > 80 or z < 20) and y < 60) or ((z > 75 or z < 25) and y < 55)) != (
            ((z > 80 or z < 20) and y < 60) or ((z > 75 and z < 25) and y < 55)): b[91] = 92
    if (((z > 80 or z < 20) and y < 60) or ((z > 75 or z < 25) and y < 55)) != (
            ((z > 80 or z < 20) and y < 60) or ((z > 75 or z != 25) and y < 55)): b[92] = 93
    if (((z > 80 or z < 20) and y < 60) or ((z > 75 or z < 25) and y < 55)) != (
            ((z > 80 or z < 20) and y < 60) or ((z > 75 or z < 35) and y < 55)): b[93] = 94
    if (((z > 80 or z < 20) and y < 60) or ((z > 75 or z < 25) and y < 55)) != (
            ((z > 80 or z < 20) and y < 60) or ((z > 75 or z < 25) or y < 55)): b[94] = 95
    if (((z > 80 or z < 20) and y < 60) or ((z > 75 or z < 25) and y < 55)) != (
            ((z > 80 or z < 20) and y < 60) or ((z > 75 or z < 25) and y != 55)): b[95] = 96
    if (((z > 80 or z < 20) and y < 60) or ((z > 75 or z < 25) and y < 55)) != (
            ((z > 80 or z < 20) and y < 60) or ((z > 75 or z < 25) and y < 39)): b[96] = 97
    if ((z > 80 or z < 20) and y < 60) or ((z > 75 or z < 25) and y < 55):
        pattern_type = 153
    if (y < 30 and x > 55) != (y < 30 or x > 55): b[97] = 98
    if (y < 30 and x > 55) != (y != 30 and x > 55): b[98] = 99
    if (y < 30 and x > 55) != (y < 19 and x > 55): b[99] = 100
    if (y < 30 and x > 55) != (y < 30 and x != 55): b[100] = 101
    if (y < 30 and x > 55) != (y < 30 and x > 45): b[101] = 102
    if y < 30 and x > 55:
        pattern_type = 154
    if ((60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50)) != (
            (60 <= x < 64 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50)): b[102] = 103
    if ((60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50)) != (
            (48 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50)): b[103] = 104
    if ((60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50)) != (
            (60 <= x < 75 or y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50)): b[104] = 105
    if ((60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50)) != (
            (60 <= x < 75 and y < 55 and (x * (39 - y)) / 100 > 22) or (x > 65 and y < 50)): b[105] = 106
    if ((60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50)) != (
            (60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x*2 > 65 and y < 50)): b[106] = 107
    if ((60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50)) != (
            (60 <= x < 75 and y < 55 or (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50)): b[107] = 108
    if ((60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50)) != (
            (60 <= x < 75 and y < 55 and (y * (60 - y)) / 100 > 22) or (x > 65 and y < 50)): b[108] = 109
    if ((60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50)) != (
            (60 <= x < 75 and y < 55 and (45 * (60 - y)) / 100 > 22) or (x > 65 and y < 50)): b[109] = 110
    if ((60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50)) != (
            (60 <= x < 75 and y < 55 and (x * (73 - y)) / 100 > 22) or (x > 65 and y < 50)): b[110] = 111
    if ((60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50)) != (
            (60 <= x < 75 and y < 55 and (x * (60 - x)) / 100 > 22) or (x > 65 and y < 50)): b[111] = 112
    if ((60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50)) != (
            (60 <= x < 75 and y < 55 and (x * (60 - 40)) / 100 > 22) or (x > 65 and y < 50)): b[112] = 113
    if ((60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50)) != (
            (60 <= x < 75 and y < 55 and (x * (60 - y)) / 83 > 22) or (x > 65 and y < 50)): b[113] = 114
    if ((60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50)) != (
            (60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 != 22) or (x > 65 and y < 50)): b[114] = 115
    if ((60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50)) != (
            (60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 32) or (x > 65 and y < 50)): b[115] = 116
    if ((60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50)) != (
            (60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) and (x > 65 and y < 50)): b[116] = 117
    if ((60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50)) != (
            (60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x != 65 and y < 50)): b[117] = 118
    if ((60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50)) != (
            (60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 85 and y < 50)): b[118] = 119
    if ((60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50)) != (
            (60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 or y < 50)): b[119] = 120
    if ((60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50)) != (
            (60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y != 50)): b[120] = 121
    if ((60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50)) != (
            (60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 60)): b[121] = 122
    if (60 <= x < 75 and y < 55 and (x * (60 - y)) / 100 > 22) or (x > 65 and y < 50):
        pattern_type = 155
    if (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)) != (((y ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)): b[
        122] = 123
    if (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)) != (((z ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)): b[
        123] = 124
    if (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)) != (((x ** 2.3) / (y + z + 1) > 30) or (x > 60 and y < 60)):b[124] = 125
    if (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)) != (((33 ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)): b[
        125] = 126
    if (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)) != (((x ** 2) / (25 + z + 1) > 30) or (x > 60 and y < 60)): b[
        126] = 127
    if (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)) != (((x ** 2) / (y + 44 + 1) > 30) or (x > 60 and y < 60)): b[
        127] = 128
    if (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)) != (((x ** 2) / (x + z + 1) > 30) or (x > 60 and y < 60)): b[
        128] = 129
    if (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)) != (((x ** 2) / (z + z + 1) > 30) or (x > 60 and y < 60)): b[
        129] = 130
    if (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)) != (((x ** 2) / (y + x + 1) > 30) or (x > 60 and y < 60)): b[
        130] = 131
    if (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)) != (((x ** 2) / (y + y + 1) > 30) or (x > 60 and y < 60)): b[
        131] = 132
    if (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)) != (((x ** 2) / (y + z + 9) > 30) or (x > 60 and y < 60)): b[
        132] = 133
    if (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)) != (((x ** 2) / (y + z + 1) != 30) or (x > 60 and y < 60)): b[
        133] = 134
    if (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)) != (((x ** 2) / (y + z + 1) > 39) or (x > 60 and y < 60)): b[
        134] = 135
    if (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)) != (((x ** 2) / (y + z + 1) > 30) and (x > 60 and y < 60)): b[
        135] = 136
    if (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)) != (((x ** 2) / (y + z + 1) > 30) or (x != 60 and y < 60)): b[
        136] = 137
    if (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)) != (((x ** 2) / (y + z + 1) > 30) or (x > 73 and y < 60)): b[
        137] = 138
    if (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)) != (((x ** 2) / (y + z + 1) > 30) or (x > 60 or y < 60)): b[
        138] = 139
    if (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)) != (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y != 60)): b[
        139] = 140
    if (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)) != (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 69)): b[
        140] = 141
    if (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)) != (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 51)): b[
        141] = 142
    if (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)) != (((x ** 2) / (y + z + 1) > 20) or (x > 60 and y < 60)): b[
        142] = 143
    if (((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60)) != (((x ** 1.4) / (y + z + 1) > 30) or (x > 60 and y < 60)):b[143] = 144
    if ((x ** 2) / (y + z + 1) > 30) or (x > 60 and y < 60):
        pattern_type = 156
    if (40 <= y < 65 and x > 50) != (40 <= y < 65 or x > 50): b[144] = 145
    if (40 <= y < 65 and x > 50) != (40 <= y < 65 and x != 50): b[145] = 145
    if (40 <= y < 65 and x > 50) != (40 <= y < 65 and x > 63): b[146] = 147
    if (40 <= y < 65 and x > 50) != (40 <= y < 75 and x > 50): b[147] = 148
    if (40 <= y < 65 and x > 50) != (29 <= y < 65 and x > 50): b[148] = 149
    if 40 <= y < 65 and x > 50:
        pattern_type = 157
    if ((50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 55 and z > 65)) != (
            (50 <= x < 75 and z > 60 and (x * z) / 100 > 32) or (x > 55 and z > 65)): b[149] = 150
    if ((50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 55 and z > 65)) != (
            (43 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 55 and z > 65)): b[150] = 151
    if ((50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 55 and z > 65)) != (
            (50 <= x < 65 or z > 60 and (x * z) / 100 > 32) or (x > 55 and z > 65)): b[151] = 152
    if ((50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 55 and z > 65)) != (
            (50 <= x < 65 and z != 60 and (x * z) / 100 > 32) or (x > 55 and z > 65)): b[152] = 153
    if ((50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 55 and z > 65)) != (
            (50 <= x < 65 and z > 74 and (x * z) / 100 > 32) or (x > 55 and z > 65)): b[153] = 154
    if ((50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 55 and z > 65)) != (
            (50 <= x < 65 and z > 60 or (x * z) / 100 > 32) or (x > 55 and z > 65)): b[154] = 155
    if ((50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 55 and z > 65)) != (
            (50 <= x < 65 and z > 60 and (z * z) / 100 > 32) or (x > 55 and z > 65)): b[155] = 156
    if ((50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 55 and z > 65)) != (
            (50 <= x < 65 and z > 60 and (x * x) / 100 > 32) or (x > 55 and z > 65)): b[156] = 157
    if ((50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 55 and z > 65)) != (
            (50 <= x < 65 and z > 60 and (77 * z) / 100 > 32) or (x > 55 and z > 65)): b[157] = 158
    if ((50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 55 and z > 65)) != (
            (50 <= x < 65 and z > 60 and (x * 66) / 100 > 32) or (x > 55 and z > 65)): b[158] = 159
    if ((50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 55 and z > 65)) != (
            (50 <= x < 65 and z > 60 and (x * z) / 76 > 32) or (x > 55 and z > 65)): b[159] = 160
    if ((50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 55 and z > 65)) != (
            (50 <= x < 65 and z > 60 and (x * z) / 100 != 32) or (x > 55 and z > 65)): b[160] = 161
    if ((50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 55 and z > 65)) != (
            (50 <= x < 65 and z > 60 and (x * z) / 100 > 44) or (x > 55 and z > 65)): b[161] = 162
    if ((50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 55 and z > 65)) != (
            (50 <= x < 65 and z > 60 and (x * z) / 100 > 32) and (x > 55 and z > 65)): b[162] = 163
    if ((50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 55 and z > 65)) != (
            (50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x != 55 and z > 65)): b[163] = 164
    if ((50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 55 and z > 65)) != (
            (50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 44 and z > 65)): b[164] = 165
    if ((50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 55 and z > 65)) != (
            (50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 55 or z > 65)): b[165] = 166
    if ((50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 55 and z > 65)) != (
            (50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 55 and z != 65)): b[166] = 167
    if ((50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 55 and z > 65)) != (
            (50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 55 and z > 49)): b[167] = 168
    if (50 <= x < 65 and z > 60 and (x * z) / 100 > 32) or (x > 55 and z > 65):
        pattern_type = 158
    if (((z > 70 or z < 30) and y > 65) or ((z > 75 or z < 25) and y > 60)) != (
            ((z != 70 or z < 30) and y > 65) or ((z > 75 or z < 25) and y > 60)): b[168] = 169
    if (((z > 70 or z < 30) and y > 65) or ((z > 75 or z < 25) and y > 60)) != (
            ((z > 81 or z < 30) and y > 65) or ((z > 75 or z < 25) and y > 60)): b[169] = 170
    if (((z > 70 or z < 30) and y > 65) or ((z > 75 or z < 25) and y > 60)) != (
            ((z > 70 and z < 30) and y > 65) or ((z > 75 or z < 25) and y > 60)): b[170] = 171
    if (((z > 70 or z < 30) and y > 65) or ((z > 75 or z < 25) and y > 60)) != (
            ((z > 70 or z != 30) and y > 65) or ((z > 75 or z < 25) and y > 60)): b[171] = 172
    if (((z > 70 or z < 30) and y > 65) or ((z > 75 or z < 25) and y > 60)) != (
            ((z > 70 or z < 40) and y > 65) or ((z > 75 or z < 25) and y > 60)): b[172] = 173
    if (((z > 70 or z < 30) and y > 65) or ((z > 75 or z < 25) and y > 60)) != (
            ((z > 70 or z < 30) or y > 65) or ((z > 75 or z < 25) and y > 60)): b[173] = 174
    if (((z > 70 or z < 30) and y > 65) or ((z > 75 or z < 25) and y > 60)) != (
            ((z > 70 or z < 30) and y != 65) or ((z > 75 or z < 25) and y > 60)): b[174] = 175
    if (((z > 70 or z < 30) and y > 65) or ((z > 75 or z < 25) and y > 60)) != (
            ((z > 70 or z < 30) and y > 77) or ((z > 75 or z < 25) and y > 60)): b[175] = 176
    if (((z > 70 or z < 30) and y > 65) or ((z > 75 or z < 25) and y > 60)) != (
            ((z > 70 or z < 30) and y > 65) and ((z > 75 or z < 25) and y > 60)): b[176] = 177
    if (((z > 70 or z < 30) and y > 65) or ((z > 75 or z < 25) and y > 60)) != (
            ((z > 70 or z < 30) and y > 65) or ((z != 75 or z < 25) and y > 60)): b[177] = 178
    if (((z > 70 or z < 30) and y > 65) or ((z > 75 or z < 25) and y > 60)) != (
            ((z > 70 or z < 30) and y > 65) or ((z > 85 or z < 25) and y > 60)): b[178] = 179
    if (((z > 70 or z < 30) and y > 65) or ((z > 75 or z < 25) and y > 60)) != (
            ((z > 70 or z < 30) and y > 65) or ((z > 75 and z < 25) and y > 60)): b[179] = 180
    if (((z > 70 or z < 30) and y > 65) or ((z > 75 or z < 25) and y > 60)) != (
            ((z > 70 or z < 30) and y > 65) or ((z > 75 or z != 25) and y > 60)): b[180] = 181
    if (((z > 70 or z < 30) and y > 65) or ((z > 75 or z < 25) and y > 60)) != (
            ((z > 70 or z < 30) and y > 65) or ((z > 75 or z < 17) and y > 60)): b[181] = 182
    if (((z > 70 or z < 30) and y > 65) or ((z > 75 or z < 25) and y > 60)) != (
            ((z > 70 or z < 30) and y > 65) or ((z > 75 or z < 25) or y > 60)): b[182] = 183
    if (((z > 70 or z < 30) and y > 65) or ((z > 75 or z < 25) and y > 60)) != (
            ((z > 70 or z < 30) and y > 65) or ((z > 75 or z < 25) and y != 60)): b[183] = 184
    if (((z > 70 or z < 30) and y > 65) or ((z > 75 or z < 25) and y > 60)) != (
            ((z > 70 or z < 30) and y > 65) or ((z > 75 or z < 25) and y > 71)): b[184] = 185
    if ((z > 70 or z < 30) and y > 65) or ((z > 75 or z < 25) and y > 60):
        pattern_type = 159
    if (20 <= x <= 45 and y >= 75 and 40 <= z <= 65) != (20 <= x <= 55 and y >= 75 and 40 <= z <= 65): b[185] = 186
    if (20 <= x <= 45 and y >= 75 and 40 <= z <= 65) != (14 <= x <= 45 and y >= 75 and 40 <= z <= 65): b[186] = 187
    if (20 <= x <= 45 and y >= 75 and 40 <= z <= 65) != (20 <= x <= 45 or y >= 75 and 40 <= z <= 65): b[187] = 188
    if (20 <= x <= 45 and y >= 75 and 40 <= z <= 65) != (20 <= x <= 45 and y != 75 and 40 <= z <= 65): b[188] = 189
    if (20 <= x <= 45 and y >= 75 and 40 <= z <= 65) != (20 <= x <= 45 and y >= 85 and 40 <= z <= 65): b[189] = 190
    if (20 <= x <= 45 and y >= 75 and 40 <= z <= 65) != (20 <= x <= 45 and y >= 75 or 40 <= z <= 65): b[190] = 191
    if (20 <= x <= 45 and y >= 75 and 40 <= z <= 65) != (20 <= x <= 45 and y >= 75 and 22 <= z <= 65): b[191] = 192
    if (20 <= x <= 45 and y >= 75 and 40 <= z <= 65) != (20 <= x <= 45 and y >= 75 and 40 <= z <= 77): b[192] = 193
    if (20 <= x <= 45 and y >= 75 and 40 <= z <= 65) != (30 <= x <= 45 and y >= 75 and 40 <= z <= 65): b[193] = 194
    if (20 <= x <= 45 and y >= 75 and 40 <= z <= 65) != (20 <= x <= 45 and y >= 66 and 40 <= z <= 65): b[194] = 195
    if 20 <= x <= 45 and y >= 75 and 40 <= z <= 65:
        pattern_type = 160

    # 返回被触发的规则编号集合
    return set(b.values())

def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """"""
    generated_path = execute_validation_rules(particle[0], particle[1], particle[2])

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # : x:1-50, y:1-50, z:1-50
        self.bounds = bounds if bounds else [(2, 100), (2, 100), (2, 100)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """target pathsPSO"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(particles[i][0], particles[i][1], particles[i][2]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(gbest_particle[0], gbest_particle[1], gbest_particle[2]),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """Path PSO"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - Path ")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations")
    print(f"Path : {len(target_paths)}")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"Path {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "" if result['success'] else f"({result['best_fitness']:.3f})"
        print(f"{status} | {result['time']:.2f}s | iterations{result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f": {success_count}/{len(target_paths)} ({success_rate:.1f}%) | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs}")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations, {len(target_paths)}Path ")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"---  {run_idx}/{num_runs} ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f": {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs} runcompleted | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 1: 
    ws1 = wb.active
    ws1.title = ""
    ws1.sheet_view.showGridLines = False

    headers = ["Run", "", "", "", "Average Iterations", "(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f" {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # 2: Path 
    ws2 = wb.create_sheet(title="Path ")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Path ID", "", "", "", "Average Iterations", "Minimum Iterations", "Maximum Iterations"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"Path  {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # 3: 
    ws3 = wb.create_sheet(title="")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Path ", "", "(x,y,z)", "", "Iterations", "Path "]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"{run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # 4: target paths
    ws4 = wb.create_sheet(title="target paths")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Path ID", "target paths", ""]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"Path  {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f" : {filename}")
    print(f"{'=' * 70}")
    print(f":")
    print(f"  1.        - {len(all_results)} run")
    print(f"  2. Path        - Path ")
    print(f"  3.    -  runPath ")
    print(f"  4. target paths       - target paths")
    print(f"{'=' * 70}\n")

    return filename


def main():


    target_paths = [
        [38, 39, 40, 41, 42, 43, 44, 45, 49, 50, 53, 55, 56, 57, 59, 62, 63, 64, 67, 68, 69, 71, 74, 76, 79, 81, 82, 84,
         86, 90, 91, 93, 95, 98, 101, 102, 104, 105, 107, 108, 118, 120, 124, 125, 128, 131, 132, 134, 137, 139, 143,
         154, 157, 162, 163, 174, 175],
        [38, 39, 40, 41, 42, 43, 44, 45, 49, 50, 53, 55, 56, 57, 59, 62, 63, 67, 68, 69, 71, 76, 79, 81, 82, 84, 86, 91,
         93, 95, 98, 101, 102, 104, 105, 107, 108, 118, 120, 124, 125, 128, 131, 132, 134, 137, 139, 143, 151, 152, 155,
         164, 165, 166, 174, 175],
        [38, 39, 40, 41, 42, 43, 44, 45, 49, 50, 53, 55, 56, 57, 59, 62, 63, 67, 69, 71, 76, 79, 81, 82, 84, 86, 90, 91,
         93, 95, 98, 101, 105, 107, 108, 118, 120, 124, 125, 128, 131, 132, 134, 137, 139, 143, 151, 152, 155, 164, 165,
         166, 174, 175, 188],
        [38, 39, 40, 41, 42, 43, 44, 45, 49, 53, 55, 56, 57, 59, 62, 63, 64, 67, 68, 69, 71, 74, 76, 79, 89, 92, 98,
         101, 102, 104, 105, 107, 108, 118, 120, 123, 126, 127, 129, 130, 133, 135, 136, 144, 145, 157, 162, 163, 174,
         175, 183, 184],
        [28, 36, 46, 47, 48, 53, 55, 56, 57, 59, 62, 63, 67, 69, 71, 76, 79, 81, 82, 84, 86, 90, 91, 93, 95, 98, 101,
         102, 104, 105, 107, 108, 118, 120, 124, 125, 128, 131, 132, 134, 137, 139, 143, 151, 152, 155, 164, 165, 166,
         174, 175],
        [38, 39, 40, 41, 43, 44, 45, 49, 55, 56, 57, 59, 61, 62, 63, 67, 68, 69, 71, 74, 76, 79, 89, 92, 98, 101, 107,
         118, 120, 124, 125, 128, 131, 132, 134, 137, 139, 143, 151, 152, 155, 164, 165, 166, 174, 175, 183, 184, 188],
        [38, 39, 40, 41, 42, 43, 44, 45, 49, 50, 65, 66, 70, 73, 76, 79, 81, 82, 84, 86, 90, 91, 93, 95, 98, 102, 104,
         105, 107, 108, 118, 120, 123, 126, 129, 130, 135, 136, 144, 145, 154, 157, 162, 163, 174, 175],
        [28, 36, 46, 47, 48, 65, 66, 70, 72, 73, 76, 79, 81, 82, 84, 86, 90, 91, 93, 95, 98, 101, 102, 104, 105, 107,
         108, 118, 120, 123, 126, 127, 129, 130, 133, 135, 136, 144, 145, 154, 157, 162, 163, 174, 175],
        [9, 11, 36, 47, 53, 55, 57, 59, 62, 63, 64, 67, 69, 71, 76, 79, 81, 84, 86, 90, 93, 94, 95, 98, 101, 107, 118,
         120, 123, 124, 126, 127, 128, 129, 130, 131, 133, 135, 136, 144, 174, 175, 188],
        [9, 11, 36, 47, 53, 55, 57, 59, 62, 63, 64, 67, 69, 71, 76, 79, 81, 84, 86, 90, 93, 95, 104, 105, 107, 108, 118,
         120, 123, 126, 129, 130, 136, 144, 145, 152, 153, 155, 166, 167, 168, 191],
        [28, 29, 36, 47, 48, 65, 66, 70, 73, 76, 79, 81, 82, 84, 86, 91, 93, 95, 98, 101, 102, 107, 118, 120, 124, 125,
         128, 131, 132, 134, 137, 139, 143, 151, 152, 155, 164, 165, 166, 174, 175],
        [28, 36, 47, 48, 53, 55, 56, 57, 59, 62, 63, 67, 69, 71, 76, 79, 81, 82, 84, 86, 90, 93, 95, 98, 101, 105, 107,
         108, 118, 120, 124, 125, 128, 131, 132, 134, 137, 139, 143, 188, 189, 191],
        [9, 11, 29, 36, 47, 52, 54, 56, 60, 61, 65, 75, 76, 79, 81, 84, 86, 90, 93, 94, 95, 98, 101, 107, 118, 120, 123,
         124, 126, 127, 128, 129, 130, 131, 133, 135, 136, 144, 174, 175, 188],
        [28, 29, 32, 33, 34, 36, 37, 47, 54, 56, 57, 59, 61, 62, 63, 64, 68, 69, 70, 74, 81, 82, 84, 86, 90, 91, 93, 95,
         118, 120, 124, 134, 137, 139, 145, 146, 164, 166, 174, 175, 188],
        [1, 16, 17, 19, 20, 21, 25, 27, 28, 29, 32, 33, 34, 36, 37, 41, 42, 47, 86, 87, 95, 96, 98, 99, 120, 121, 123,
         124, 125, 127, 128, 129, 131, 132, 134, 139, 140, 143, 145, 163],
        [1, 19, 25, 27, 29, 32, 36, 37, 47, 98, 99, 105, 123, 125, 127, 128, 129, 130, 134, 139, 140, 143, 145, 148,
         152, 153, 155, 166, 167, 168, 169, 172, 174, 178, 181, 183, 191],
        [38, 39, 40, 41, 43, 44, 45, 49, 50, 65, 66, 70, 73, 76, 79, 89, 92, 100, 104, 105, 107, 108, 118, 120, 123,
         126, 127, 129, 130, 133, 135, 136, 144, 145, 174, 175, 183, 184],
        [9, 11, 31, 36, 47, 52, 54, 56, 58, 60, 61, 65, 75, 76, 79, 89, 92, 98, 101, 107, 118, 120, 123, 124, 127, 128,
         129, 130, 131, 133, 135, 136, 144, 174, 175, 183, 184, 188],
        [6, 9, 11, 29, 36, 47, 76, 79, 81, 84, 86, 90, 93, 95, 107, 118, 120, 124, 125, 127, 128, 131, 132, 134, 137,
         139, 143, 145, 149, 152, 155, 156, 158, 159, 160, 161, 191],
        [19, 24, 25, 27, 29, 32, 36, 37, 47, 98, 99, 105, 120, 121, 124, 126, 132, 133, 135, 136, 144, 145, 150, 152,
         155, 166, 167, 168, 169, 172, 174, 178, 181, 183, 188, 191],
        [27, 28, 29, 32, 33, 34, 36, 37, 40, 42, 46, 47, 48, 54, 56, 57, 59, 61, 62, 63, 64, 68, 69, 70, 83, 88, 89,
         124, 134, 137, 139, 145, 146, 164, 166, 174, 175, 183, 184],
        [1, 13, 16, 17, 19, 20, 21, 25, 27, 28, 29, 32, 36, 37, 42, 47, 86, 87, 95, 96, 98, 99, 105, 120, 121, 123, 124,
         125, 127, 128, 131, 132, 134, 139, 140, 143, 145, 163],
        [38, 40, 41, 43, 45, 53, 55, 56, 57, 59, 62, 63, 64, 67, 68, 69, 71, 76, 79, 106, 109, 112, 113, 116, 117, 123,
         126, 129, 130, 135, 136, 144, 145, 174, 175, 183, 184],
        [19, 26, 28, 30, 31, 34, 38, 98, 99, 105, 123, 125, 127, 128, 129, 130, 134, 139, 140, 143, 145, 152, 153, 155,
         166, 167, 168, 169, 172, 174, 178, 181, 183, 188, 191],
        [1, 3, 8, 13, 16, 17, 19, 20, 21, 27, 28, 29, 36, 42, 47, 86, 87, 95, 96, 98, 99, 105, 120, 121, 123, 124, 125,
         127, 128, 131, 132, 134, 139, 140, 143, 145, 148, 163],
        [2, 4, 5, 8, 19, 24, 27, 29, 36, 47, 76, 79, 89, 92, 97, 98, 99, 105, 107, 108, 115, 118, 120, 123, 124, 126,
         136, 144, 145, 149, 152, 166, 167, 174, 175, 183, 184],
        [9, 11, 29, 36, 47, 76, 79, 81, 84, 86, 90, 93, 95, 98, 99, 105, 107, 108, 111, 115, 118, 120, 123, 126, 135,
         136, 144, 145, 149, 152, 153, 155, 166, 167, 168, 191],
        [1, 3, 7, 9, 10, 11, 24, 29, 31, 36, 47, 60, 61, 65, 75, 76, 79, 98, 101, 118, 120, 123, 124, 127, 128, 129,
         131, 132, 133, 135, 136, 144, 174, 175, 183, 184, 188],
        [2, 3, 5, 6, 9, 10, 11, 27, 29, 36, 47, 81, 84, 86, 90, 93, 95, 98, 99, 105, 108, 115, 123, 126, 129, 130, 133,
         135, 136, 144, 147, 152, 155, 166, 167, 168, 191],
        [19, 26, 28, 30, 31, 33, 34, 35, 38, 98, 99, 105, 120, 121, 124, 126, 135, 136, 144, 145, 150, 152, 155, 166,
         167, 168, 169, 172, 174, 178, 181, 183, 188, 191],
        [2, 3, 4, 5, 6, 9, 10, 11, 27, 28, 29, 36, 42, 47, 86, 87, 95, 96, 98, 99, 105, 120, 121, 124, 125, 127, 128,
         131, 132, 134, 139, 140, 141, 143, 145, 148, 163],
        [2, 4, 5, 6, 8, 22, 24, 27, 29, 31, 36, 47, 60, 65, 76, 79, 98, 101, 118, 120, 123, 124, 127, 128, 129, 131,
         132, 133, 135, 136, 144, 174, 175, 183, 184, 188],
        [9, 11, 36, 47, 48, 53, 55, 57, 59, 62, 63, 64, 67, 69, 71, 76, 79, 81, 84, 86, 90, 93, 95, 103, 109, 112, 113,
         117, 145, 150, 152, 155, 166, 167, 168, 191],
        [38, 39, 40, 41, 43, 44, 45, 76, 79, 100, 105, 107, 108, 111, 114, 115, 118, 120, 124, 125, 127, 128, 131, 132,
         134, 139, 143, 145, 149, 174, 175, 183, 184],
        [11, 12, 14, 15, 18, 22, 23, 27, 28, 29, 32, 36, 37, 42, 47, 86, 87, 95, 96, 98, 99, 105, 123, 124, 125, 127,
         128, 131, 134, 139, 140, 143, 145, 148, 163],
        [19, 24, 25, 27, 29, 32, 36, 37, 45, 47, 76, 77, 98, 99, 120, 121, 124, 126, 136, 144, 145, 155, 166, 167, 168,
         169, 172, 174, 178, 181, 183, 188, 191],
        [2, 4, 5, 6, 9, 10, 11, 27, 28, 29, 36, 42, 47, 86, 87, 95, 96, 98, 99, 105, 124, 125, 127, 128, 131, 134, 139,
         140, 141, 143, 145, 148, 177, 180, 185],
        [2, 3, 4, 5, 6, 9, 10, 11, 27, 29, 36, 42, 47, 86, 87, 95, 96, 98, 99, 105, 120, 121, 123, 126, 129, 130, 133,
         135, 136, 144, 163, 177, 179, 180, 185],
        [1, 16, 19, 25, 27, 29, 32, 36, 37, 47, 98, 99, 105, 120, 121, 123, 124, 125, 127, 128, 129, 130, 131, 134, 139,
         140, 143, 145, 163, 170, 171, 177],
        [1, 16, 19, 21, 25, 27, 29, 36, 37, 47, 98, 99, 105, 123, 124, 125, 127, 128, 129, 131, 134, 139, 140, 143, 145,
         148, 163, 170, 171, 176, 177],
        [19, 26, 28, 30, 31, 34, 38, 98, 99, 124, 126, 128, 131, 132, 133, 135, 136, 144, 145, 148, 152, 166, 167, 169,
         172, 173, 174, 178, 181, 183],
        [4, 9, 11, 29, 31, 33, 36, 37, 47, 51, 52, 54, 56, 59, 61, 76, 79, 81, 84, 86, 90, 93, 94, 95, 98, 101, 118,
         120, 134, 137, 139, 174, 175],
        [1, 3, 9, 10, 11, 29, 36, 47, 76, 79, 100, 106, 109, 110, 112, 113, 116, 117, 123, 124, 136, 144, 145, 152, 166,
         167, 174, 175, 183, 184],
        [2, 3, 5, 6, 9, 10, 11, 27, 29, 36, 47, 81, 84, 86, 90, 93, 95, 98, 99, 105, 108, 115, 120, 121, 122, 150, 152,
         155, 166, 167, 168, 191],
        [4, 24, 25, 27, 29, 36, 44, 47, 76, 77, 86, 87, 95, 96, 98, 99, 120, 121, 124, 126, 136, 144, 145, 148, 166,
         167, 177, 180, 182, 185],
        [9, 11, 36, 39, 44, 46, 47, 48, 52, 54, 56, 60, 65, 75, 80, 81, 82, 84, 86, 90, 93, 95, 117, 119, 145, 152, 155,
         166, 167, 168, 191],
        [19, 25, 26, 28, 30, 31, 33, 34, 38, 83, 85, 88, 89, 123, 125, 127, 129, 130, 134, 137, 139, 143, 145, 146, 174,
         175, 183, 184, 188],
        [16, 19, 21, 26, 28, 30, 31, 38, 123, 125, 134, 152, 155, 156, 158, 159, 160, 161, 169, 172, 174, 178, 181, 183,
         186, 188, 191],
        [2, 4, 5, 6, 9, 10, 11, 27, 28, 29, 36, 42, 47, 82, 84, 86, 95, 96, 98, 99, 105, 136, 138, 142, 147, 174, 175,
         183, 184],
        [1, 3, 13, 16, 19, 21, 24, 30, 31, 38, 53, 55, 60, 65, 75, 123, 124, 134, 164, 166, 170, 171, 176, 177, 188,
         191, 193],
        [1, 3, 9, 10, 11, 29, 36, 47, 78, 80, 81, 84, 86, 90, 93, 94, 95, 98, 99, 117, 119, 145, 149, 166, 167, 174,
         175],
        [3, 13, 16, 19, 21, 28, 30, 38, 53, 55, 60, 65, 75, 123, 134, 169, 172, 174, 178, 181, 183, 188, 189, 191, 195],
        [3, 13, 16, 19, 21, 38, 53, 55, 57, 60, 65, 75, 123, 134, 169, 172, 174, 178, 181, 183, 187, 188, 191],
        [13, 16, 19, 28, 30, 38, 55, 60, 65, 75, 123, 134, 169, 173, 174, 178, 181, 183, 188, 191, 192],
        [1, 3, 13, 16, 19, 21, 24, 30, 31, 38, 123, 134, 169, 172, 174, 178, 181, 183, 189, 190, 194]
    ]

    print("=" * 70)
    print("baseline PSO")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("Program completed")


if __name__ == "__main__":
    main()