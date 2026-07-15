import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import math

def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

def execute_validation_rules(a):
    """
    替换原有的编排规则函数
    参数a: 包含3个元素的元组或数组，分别对应path_depth, file_count, access_level
    返回: 触发的规则编号集合
    """
    path_depth, file_count, access_level = float(a[0]), int(a[1]), float(a[2])

    # 使用正确的变量名
    x, y, z = path_depth, file_count, access_level

    triggered = set()
    # 创建一个字典来存储b数组的值，用于跟踪哪些规则被触发
    b = {}
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2): b[0] = 1
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((27 * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): b[1] = 2
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * 63) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): b[2] = 3
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (36 + 1) > 50 and x ** 2 + y ** 2 > z ** 2): b[3] = 4
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 7) > 50 and x ** 2 + y ** 2 > z ** 2): b[4] = 5
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 12) > 50 and x ** 2 + y ** 2 > z ** 2): b[5] = 6
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): b[6] = 7
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((z * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): b[7] = 8
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * x) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): b[8] = 9
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): b[9] = 10
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (x + 1) > 50 and x ** 2 + y ** 2 > z ** 2): b[10] = 11
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (y + 1) > 50 and x ** 2 + y ** 2 > z ** 2): b[11] = 12
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) != 50 and x ** 2 + y ** 2 > z ** 2): b[12] = 13
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 70 and x ** 2 + y ** 2 > z ** 2): b[13] = 14
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 28 and x ** 2 + y ** 2 > z ** 2): b[14] = 15
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and y ** 2 + y ** 2 > z ** 2): b[15] = 16
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 1 + y ** 2 > z ** 2): b[16] = 17
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 1 > z ** 2): b[17] = 18
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + x ** 2 > z ** 2): b[18] = 19
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 3): b[19] = 20
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x * 2 + y ** 2 > z ** 2): b[20] = 21
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2.5): b[21] = 22
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z * 1) > 50 and x ** 2 + y ** 2 > z ** 2): b[22] = 23
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z - 1) > 50 and x ** 2 + y ** 2 > z ** 2): b[23] = 24
    if (x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2:
        pattern_type = "291A1"
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 3 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[24] = 25
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 4 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[25] = 26
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - z ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[26] = 27
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 1) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[27] = 28
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[28] = 29
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 1 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[29] = 30
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (x + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[30] = 31
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[31] = 32
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (z + 1.5) < -30 or (abs(x - y) * z) / 100 > 45): b[32] = 33
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (z - 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[33] = 34
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (z + 0.1) != -30 or (abs(x - y) * z) / 100 > 45): b[34] = 35
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (z + 0.1) < -10 or (abs(x - y) * z) / 100 > 45): b[35] = 36
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (z + 0.1) < -30 and (abs(x - y) * z) / 100 > 45): b[36] = 37
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((24 ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[37] = 38
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - 19 ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[38] = 39
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (42 + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[39] = 40
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45): b[40] = 41
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - z) * z) / 100 > 45): b[41] = 42
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(52 - y) * z) / 100 > 45): b[42] = 43
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - 42) * z) / 100 > 45): b[43] = 44
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * 55) / 100 > 45): b[44] = 45
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 75 > 45): b[45] = 46
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 25): b[46] = 47
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 65): b[47] = 48
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 != 45): b[48] = 49
    if (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45:
        pattern_type = "291A2"
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((46 + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35): b[49] = 50
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((x + 37) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35): b[50] = 51
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((x + y) / 3) ** 2 < z * 20 and (x * y * z) / 1000 < 35): b[51] = 52
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((x + y) / 2) ** 3 < z * 20 and (x * y * z) / 1000 < 35): b[52] = 53
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((x + y) / 2) ** 2 != z * 20 and (x * y * z) / 1000 < 35): b[53] = 54
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((x + y) / 2) ** 2 < x * 20 and (x * y * z) / 1000 < 35): b[54] = 55
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35): b[55] = 56
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 30 and (x * y * z) / 1000 < 35): b[56] = 57
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 or (x * y * z) / 1000 < 35): b[57] = 58
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((x + y) / 2) ** 2 < 52 * 20 and (x * y * z) / 1000 < 35): b[58] = 59
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((x + y) / 2) ** 3 < z * 20 and (x * y * z) / 1000 < 35): b[59] = 60
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (y * y * z) / 1000 < 35): b[60] = 61
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (z * y * z) / 1000 < 35): b[61] = 62
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35): b[62] = 63
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * z * z) / 1000 < 35): b[63] = 64
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35): b[64] = 65
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * y * y) / 1000 < 35): b[65] = 66
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1200 < 35): b[66] = 67
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 700 < 35): b[67] = 68
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 15): b[68] = 69
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 55): b[69] = 70
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 != 35): b[70] = 71
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (34 * y * z) / 1000 < 35): b[71] = 72
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * 53 * z) / 1000 < 35): b[72] = 73
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * y * 72) / 1000 < 35): b[73] = 74
    if ((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35:
        pattern_type = "291A3"
    if (x < 50 and y > 80) != (x < 50 or y > 80): b[74] = 75
    if (x < 50 and y > 80) != (x < 50 and y != 80): b[75] = 76
    if (x < 50 and y > 80) != (x != 50 and y > 80): b[76] = 77
    if (x < 50 and y > 80) != (x < 37 and y > 80): b[77] = 78
    if (x < 50 and y > 80) != (x < 50 and y > 89): b[78] = 79
    if x < 50 and y > 80:
        pattern_type = 302
    if ((y > 92 and z < 40) or (y > 90 and z < 35 and x > 65)) != ((y > 92 or z < 40) or (y > 90 and z < 35 and x > 65)): b[79] = 80
    if ((y > 92 and z < 40) or (y > 90 and z < 35 and x > 65)) != ((y != 92 and z < 40) or (y > 90 and z < 35 and x > 65)): b[80] = 81
    if ((y > 92 and z < 40) or (y > 90 and z < 35 and x > 65)) != ((y > 92 and z != 40) or (y > 90 and z < 35 and x > 65)): b[81] = 82
    if ((y > 92 and z < 40) or (y > 90 and z < 35 and x > 65)) != ((y > 92 and z < 51) or (y > 90 and z < 35 and x > 65)): b[82] = 83
    if ((y > 92 and z < 40) or (y > 90 and z < 35 and x > 65)) != ((y > 82 and z < 40) or (y > 90 and z < 35 and x > 65)): b[83] = 84
    if ((y > 92 and z < 40) or (y > 90 and z < 35 and x > 65)) != ((y > 92 and z < 40) and (y > 90 and z < 35 and x > 65)): b[84] = 85
    if ((y > 92 and z < 40) or (y > 90 and z < 35 and x > 65)) != ((y > 92 and z < 40) or (y > 90 or z < 35 and x > 65)): b[85] = 86
    if ((y > 92 and z < 40) or (y > 90 and z < 35 and x > 65)) != ((y > 92 and z < 40) or (y > 90 and z < 35 or x > 65)): b[86] = 87
    if ((y > 92 and z < 40) or (y > 90 and z < 35 and x > 65)) != ((y > 92 and z < 40) or (y != 90 and z < 35 and x > 65)): b[87] = 88
    if ((y > 92 and z < 40) or (y > 90 and z < 35 and x > 65)) != ((y > 92 and z < 40) or (y > 90 and z != 35 and x > 65)): b[88] = 89
    if ((y > 92 and z < 40) or (y > 90 and z < 35 and x > 65)) != ((y > 92 and z < 40) or (y > 90 and z < 35 and x != 65)): b[89] = 90
    if ((y > 92 and z < 40) or (y > 90 and z < 35 and x > 65)) != ((y > 92 and z < 40) or (y > 78 and z < 35 and x > 65)): b[90] = 91
    if ((y > 92 and z < 40) or (y > 90 and z < 35 and x > 65)) != ((y > 92 and z < 40) or (y > 90 and z < 44 and x > 65)): b[91] = 92
    if ((y > 92 and z < 40) or (y > 90 and z < 35 and x > 65)) != ((y > 92 and z < 40) or (y > 90 and z < 35 and x > 81)): b[92] = 93
    if (y > 92 and z < 40) or (y > 90 and z < 35 and x > 65):
        pattern_type = 303
    if (z < 25 and x > 60) != (z < 25 or x > 60): b[93] = 94
    if (z < 25 and x > 60) != (z != 25 and x > 60): b[94] = 95
    if (z < 25 and x > 60) != (z < 25 and x != 60): b[95] = 96
    if (z < 25 and x > 60) != (z < 25 and x > 69): b[96] = 97
    if (z < 25 and x > 60) != (z < 33 and x > 60): b[97] = 98
    if z < 25 and x > 60:
        pattern_type = 304
    if ((50 <= x < 75 and y > 70 and (x * y) / 100 < 58) or (x < 70 and y > 75)) != ((50 <= x < 75 or y > 70 and (x * y) / 100 < 58) or (x < 70 and y > 75)): b[98] = 99
    if ((50 <= x < 75 and y > 70 and (x * y) / 100 < 58) or (x < 70 and y > 75)) != ((50 <= x < 75 and y > 70 or (x * y) / 100 < 58) or (x < 70 and y > 75)): b[99] = 100
    if ((50 <= x < 75 and y > 70 and (x * y) / 100 < 58) or (x < 70 and y > 75)) != ((50 <= x < 75 and y > 70 and (x * y) / 100 < 58) and (x < 70 and y > 75)): b[100] = 101
    if ((50 <= x < 75 and y > 70 and (x * y) / 100 < 58) or (x < 70 and y > 75)) != ((50 <= x < 81 and y > 70 and (x * y) / 100 < 58) or (x < 70 and y > 75)): b[101] = 102
    if ((50 <= x < 75 and y > 70 and (x * y) / 100 < 58) or (x < 70 and y > 75)) != ((33 <= x < 75 and y > 70 and (x * y) / 100 < 58) or (x < 70 and y > 75)): b[102] = 103
    if ((50 <= x < 75 and y > 70 and (x * y) / 100 < 58) or (x < 70 and y > 75)) != ((50 <= x < 75 and y != 70 and (x * y) / 100 < 58) or (x < 70 and y > 75)): b[103] = 104
    if ((50 <= x < 75 and y > 70 and (x * y) / 100 < 58) or (x < 70 and y > 75)) != ((50 <= x < 75 and y > 79 and (x * y) / 100 < 58) or (x < 70 and y > 75)): b[104] = 105
    if ((50 <= x < 75 and y > 70 and (x * y) / 100 < 58) or (x < 70 and y > 75)) != ((50 <= x < 75 and y > 70 and (y * y) / 100 < 58) or (x < 70 and y > 75)): b[105] = 106
    if ((50 <= x < 75 and y > 70 and (x * y) / 100 < 58) or (x < 70 and y > 75)) != ((50 <= x < 75 and y > 70 and (x * x) / 100 < 58) or (x < 70 and y > 75)): b[106] = 107
    if ((50 <= x < 75 and y > 70 and (x * y) / 100 < 58) or (x < 70 and y > 75)) != ((50 <= x < 75 and y > 70 and (x * y) / 87 < 58) or (x < 70 and y > 75)): b[107] = 108
    if ((50 <= x < 75 and y > 70 and (x * y) / 100 < 58) or (x < 70 and y > 75)) != ((50 <= x < 75 and y > 70 and (x * y) / 100 != 58) or (x < 70 and y > 75)): b[108] = 109
    if ((50 <= x < 75 and y > 70 and (x * y) / 100 < 58) or (x < 70 and y > 75)) != ((50 <= x < 75 and y > 70 and (x * y) / 100 < 68) or (x < 70 and y > 75)): b[109] = 110
    if ((50 <= x < 75 and y > 70 and (x * y) / 100 < 58) or (x < 70 and y > 75)) != ((50 <= x < 75 and y > 70 and (x * y) / 100 < 58) or (x != 70 and y > 75)): b[110] = 111
    if ((50 <= x < 75 and y > 70 and (x * y) / 100 < 58) or (x < 70 and y > 75)) != ((50 <= x < 75 and y > 70 and (x * y) / 100 < 58) or (x < 70 and y != 75)): b[111] = 112
    if ((50 <= x < 75 and y > 70 and (x * y) / 100 < 58) or (x < 70 and y > 75)) != ((50 <= x < 75 and y > 70 and (x * y) / 100 < 58) or (x < 70 and y > 65)): b[112] = 113
    if (50 <= x < 75 and y > 70 and (x * y) / 100 < 58) or (x < 70 and y > 75):
        pattern_type = 305
    if (((z ** 2) / (x + 100 - y + 1) < 15) or (z < 50 and x < 75)) != (((z ** 2) / (x + 100 - y + 1) < 15) and (z < 50 and x < 75)): b[113] = 114
    if (((z ** 2) / (x + 100 - y + 1) < 15) or (z < 50 and x < 75)) != (((z ** 2) / (x + 100 - y + 1) < 15) or (z < 50 or x < 75)): b[114] = 115
    if (((z ** 2) / (x + 100 - y + 1) < 15) or (z < 50 and x < 75)) != (((z ** 2) / (x + 100 - y + 1) < 15) or (z != 50 and x < 75)): b[115] = 116
    if (((z ** 2) / (x + 100 - y + 1) < 15) or (z < 50 and x < 75)) != (((z ** 2) / (x + 100 - y + 1) < 15) or (z < 40 and x < 75)): b[116] = 117
    if (((z ** 2) / (x + 100 - y + 1) < 15) or (z < 50 and x < 75)) != (((z ** 2) / (x + 100 - y + 1) < 15) or (z < 50 and x != 75)): b[117] = 118
    if (((z ** 2) / (x + 100 - y + 1) < 15) or (z < 50 and x < 75)) != (((z ** 2) / (x + 100 - y + 1) < 15) or (z < 50 and x < 86)): b[118] = 119
    if (((z ** 2) / (x + 100 - y + 1) < 15) or (z < 50 and x < 75)) != (((z ** 2) / (x + 100 - y + 1) != 15) or (z < 50 and x < 75)): b[119] = 120
    if (((z ** 2) / (x + 100 - y + 1) < 15) or (z < 50 and x < 75)) != (((z ** 2) / (x + 100 - y + 1) < 25) or (z < 50 and x < 75)): b[120] = 121
    if (((z ** 2) / (x + 100 - y + 1) < 15) or (z < 50 and x < 75)) != (((z ** 2) / (x + 100 - z + 1) < 15) or (z < 50 and x < 75)): b[121] = 122
    if (((z ** 2) / (x + 100 - y + 1) < 15) or (z < 50 and x < 75)) != (((z ** 2) / (z + 100 - y + 1) < 15) or (z < 50 and x < 75)): b[122] = 123
    if (((z ** 2) / (x + 100 - y + 1) < 15) or (z < 50 and x < 75)) != (((x ** 2) / (x + 100 - y + 1) < 15) or (z < 50 and x < 75)): b[123] = 124
    if (((z ** 2) / (x + 100 - y + 1) < 15) or (z < 50 and x < 75)) != (((y ** 2) / (x + 100 - y + 1) < 15) or (z < 50 and x < 75)): b[124] = 125
    if (((z ** 2) / (x + 100 - y + 1) < 15) or (z < 50 and x < 75)) != (((54 ** 2) / (x + 100 - y + 1) < 15) or (z < 50 and x < 75)): b[125] = 126
    if (((z ** 2) / (x + 100 - y + 1) < 15) or (z < 50 and x < 75)) != (((z ** 2) / (73 + 100 - y + 1) < 15) or (z < 50 and x < 75)): b[126] = 127
    if (((z ** 2) / (x + 100 - y + 1) < 15) or (z < 50 and x < 75)) != (((z ** 2) / (x + 100 - 26 + 1) < 15) or (z < 50 and x < 75)): b[127] = 128
    if (((z ** 2) / (x + 100 - y + 1) < 15) or (z < 50 and x < 75)) != (((z ** 2) / (x + 100 - y + 10) < 15) or (z < 50 and x < 75)): b[128] = 129
    if (((z ** 2) / (x + 100 - y + 1) < 15) or (z < 50 and x < 75)) != (((z ** 2) / (x + 113 - y + 1) < 15) or (z < 50 and x < 75)): b[129] = 130
    if ((z ** 2) / (x + 100 - y + 1) < 15) or (z < 50 and x < 75):
        pattern_type = 306
    if (80 <= y < 90 and x > 70) != (80 <= y < 90 or x > 70): b[130] = 131
    if (80 <= y < 90 and x > 70) != (80 <= y < 90 and x != 70): b[131] = 132
    if (80 <= y < 90 and x > 70) != (80 <= y < 90 and x > 55): b[132] = 133
    if (80 <= y < 90 and x > 70) != (80 <= y < 96 and x > 70): b[133] = 134
    if (80 <= y < 90 and x > 70) != (71 <= y < 90 and x > 70): b[134] = 135
    if 80 <= y < 90 and x > 70:
        pattern_type = 307
    if ((40 <= z < 65 and x > 75 and (z * x) / 100 < 52) or (z < 60 and x > 80)) != ((40 <= z < 65 or x > 75 and (z * x) / 100 < 52) or (z < 60 and x > 80)): b[135] = 136
    if ((40 <= z < 65 and x > 75 and (z * x) / 100 < 52) or (z < 60 and x > 80)) != ((40 <= z < 65 and x > 75 or (z * x) / 100 < 52) or (z < 60 and x > 80)): b[136] = 137
    if ((40 <= z < 65 and x > 75 and (z * x) / 100 < 52) or (z < 60 and x > 80)) != ((40 <= z < 65 and x > 75 and (z * x) / 100 < 52) and (z < 60 and x > 80)): b[137] = 138
    if ((40 <= z < 65 and x > 75 and (z * x) / 100 < 52) or (z < 60 and x > 80)) != ((40 <= z < 65 and x > 75 and (z * x) / 100 < 52) or (z != 60 and x > 80)): b[138] = 139
    if ((40 <= z < 65 and x > 75 and (z * x) / 100 < 52) or (z < 60 and x > 80)) != ((40 <= z < 65 and x > 75 and (z * x) / 100 < 52) or (z < 60 or x > 80)): b[139] = 140
    if ((40 <= z < 65 and x > 75 and (z * x) / 100 < 52) or (z < 60 and x > 80)) != ((40 <= z < 65 and x > 75 and (z * x) / 100 < 52) or (z < 60 and x != 80)): b[140] = 141
    if ((40 <= z < 65 and x > 75 and (z * x) / 100 < 52) or (z < 60 and x > 80)) != ((40 <= z < 65 and x > 75 and (z * x) / 100 < 52) or (z < 60 and x > 65)): b[141] = 142
    if ((40 <= z < 65 and x > 75 and (z * x) / 100 < 52) or (z < 60 and x > 80)) != ((40 <= z < 65 and x > 75 and (z * x) / 100 < 52) or (z < 77 and x > 80)): b[142] = 143
    if ((40 <= z < 65 and x > 75 and (z * x) / 100 < 52) or (z < 60 and x > 80)) != ((40 <= z < 65 and x > 75 and (z * x) / 100 < 66) or (z < 60 and x > 80)): b[143] = 144
    if ((40 <= z < 65 and x > 75 and (z * x) / 100 < 52) or (z < 60 and x > 80)) != ((40 <= z < 65 and x > 75 and (z * z) / 100 < 52) or (z < 60 and x > 80)): b[144] = 145
    if ((40 <= z < 65 and x > 75 and (z * x) / 100 < 52) or (z < 60 and x > 80)) != ((40 <= z < 65 and x > 75 and (x * x) / 100 < 52) or (z < 60 and x > 80)): b[145] = 146
    if ((40 <= z < 65 and x > 75 and (z * x) / 100 < 52) or (z < 60 and x > 80)) != ((40 <= z < 65 and x > 61 and (z * x) / 100 < 52) or (z < 60 and x > 80)): b[146] = 147
    if ((40 <= z < 65 and x > 75 and (z * x) / 100 < 52) or (z < 60 and x > 80)) != ((40 <= z < 65 and x != 75 and (z * x) / 100 < 52) or (z < 60 and x > 80)): b[147] = 148
    if (40 <= z < 65 and x > 75 and (z * x) / 100 < 52) or (z < 60 and x > 80):
        pattern_type = 308
    if ((65 <= y < 78 and z > 60) or (70 <= y < 82 and z > 55)) != ((65 <= y < 78 or z > 60) or (70 <= y < 82 and z > 55)): b[148] = 149
    if ((65 <= y < 78 and z > 60) or (70 <= y < 82 and z > 55)) != ((65 <= y < 78 and z != 60) or (70 <= y < 82 and z > 55)): b[149] = 150
    if ((65 <= y < 78 and z > 60) or (70 <= y < 82 and z > 55)) != ((65 <= y < 78 and z > 60) and (70 <= y < 82 and z > 55)): b[150] = 151
    if ((65 <= y < 78 and z > 60) or (70 <= y < 82 and z > 55)) != ((65 <= y < 78 and z > 60) or (70 <= y < 82 or z > 55)): b[151] = 152
    if ((65 <= y < 78 and z > 60) or (70 <= y < 82 and z > 55)) != ((65 <= y < 78 and z > 60) or (70 <= y < 82 and z != 55)): b[152] = 153
    if ((65 <= y < 78 and z > 60) or (70 <= y < 82 and z > 55)) != ((65 <= y < 78 and z > 60) or (70 <= y < 82 and z > 75)): b[153] = 154
    if ((65 <= y < 78 and z > 60) or (70 <= y < 82 and z > 55)) != ((65 <= y < 78 and z > 60) or (70 <= y < 89 and z > 55)): b[154] = 155
    if ((65 <= y < 78 and z > 60) or (70 <= y < 82 and z > 55)) != ((65 <= y < 78 and z > 60) or (55 <= y < 82 and z > 55)): b[155] = 156
    if ((65 <= y < 78 and z > 60) or (70 <= y < 82 and z > 55)) != ((65 <= y < 78 and z > 69) or (70 <= y < 82 and z > 55)): b[156] = 157
    if (65 <= y < 78 and z > 60) or (70 <= y < 82 and z > 55):
        pattern_type = 309
    if (x >= 90 and 30 <= y <= 60 and z >= 75) != (x >= 90 or 30 <= y <= 60 and z >= 75): b[157] = 158
    if (x >= 90 and 30 <= y <= 60 and z >= 75) != (x >= 90 and 30 <= y <= 60 or z >= 75): b[158] = 159
    if (x >= 90 and 30 <= y <= 60 and z >= 75) != (x >= 90 and 30 <= y <= 60 and z != 75): b[159] = 160
    if (x >= 90 and 30 <= y <= 60 and z >= 75) != (x >= 90 and 30 <= y <= 60 and z >= 88): b[160] = 161
    if (x >= 90 and 30 <= y <= 60 and z >= 75) != (x >= 90 and 30 <= y <= 69 and z >= 75): b[161] = 162
    if (x >= 90 and 30 <= y <= 60 and z >= 75) != (x >= 90 and 22 <= y <= 60 and z >= 75): b[162] = 163
    if (x >= 90 and 30 <= y <= 60 and z >= 75) != (x != 90 and 30 <= y <= 60 and z >= 75): b[163] = 164
    if (x >= 90 and 30 <= y <= 60 and z >= 75) != (x >= 72 and 30 <= y <= 60 and z >= 75): b[164] = 165
    if x >= 90 and 30 <= y <= 60 and z >= 75:
        pattern_type = 310

    # 返回被触发的规则编号集合
    return set(b.values())


def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """"""
    generated_path = execute_validation_rules(particle[0], particle[1], particle[2])

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # : x:1-50, y:1-50, z:1-50
        self.bounds = bounds if bounds else [(1, 100), (1, 100), (1, 100)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """target pathsPSO"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(particles[i][0], particles[i][1], particles[i][2]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(gbest_particle[0], gbest_particle[1], gbest_particle[2]),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """Path PSO"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - Path ")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations")
    print(f"Path : {len(target_paths)}")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"Path {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "" if result['success'] else f"({result['best_fitness']:.3f})"
        print(f"{status} | {result['time']:.2f}s | iterations{result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f": {success_count}/{len(target_paths)} ({success_rate:.1f}%) | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs}")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations, {len(target_paths)}Path ")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"---  {run_idx}/{num_runs} ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f": {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs} runcompleted | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 1: 
    ws1 = wb.active
    ws1.title = ""
    ws1.sheet_view.showGridLines = False

    headers = ["Run", "", "", "", "Average Iterations", "(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f" {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # 2: Path 
    ws2 = wb.create_sheet(title="Path ")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Path ID", "", "", "", "Average Iterations", "Minimum Iterations", "Maximum Iterations"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"Path  {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # 3: 
    ws3 = wb.create_sheet(title="")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Path ", "", "(x,y,z)", "", "Iterations", "Path "]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"{run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # 4: target paths
    ws4 = wb.create_sheet(title="target paths")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Path ID", "target paths", ""]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"Path  {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f" : {filename}")
    print(f"{'=' * 70}")
    print(f":")
    print(f"  1.        - {len(all_results)} run")
    print(f"  2. Path        - Path ")
    print(f"  3.    -  runPath ")
    print(f"  4. target paths       - target paths")
    print(f"{'=' * 70}\n")

    return filename


def main():


    target_paths = [
        {20, 29, 30, 35, 38, 49, 80, 81, 87, 94, 95, 99, 100, 102, 111, 115, 118, 119, 120, 121, 122, 128, 129, 130,
         131, 135, 136, 137, 140, 141, 142, 149, 150, 152, 153},
        {2, 3, 14, 17, 18, 20, 21, 22, 27, 30, 31, 32, 35, 36, 38, 40, 49, 75, 77, 86, 87, 89, 94, 95, 99, 100, 107,
         109, 110, 111, 115, 116, 120, 131, 134, 149, 152, 159},
        {2, 3, 14, 17, 18, 20, 21, 22, 25, 26, 28, 29, 37, 39, 75, 77, 80, 82, 86, 87, 89, 94, 95, 99, 100, 107, 109,
         110, 111, 115, 116, 120, 131, 134, 149, 152, 159},
        {1, 3, 9, 10, 12, 13, 35, 46, 47, 49, 50, 52, 54, 55, 57, 58, 87, 94, 95, 99, 100, 104, 114, 122, 123, 124, 126,
         128, 131, 136, 137, 140, 141, 142, 147, 148},
        {1, 3, 9, 10, 12, 13, 35, 47, 49, 50, 52, 54, 57, 58, 87, 94, 95, 99, 100, 104, 115, 116, 120, 121, 125, 129,
         130, 131, 136, 137, 140, 141, 142, 147, 148},
        {20, 25, 26, 27, 28, 37, 39, 75, 77, 80, 83, 86, 87, 89, 92, 94, 95, 99, 100, 107, 109, 110, 111, 114, 117, 131,
         134, 136, 137, 140, 141, 142, 147, 148},
        {27, 31, 32, 35, 36, 40, 41, 42, 47, 49, 50, 53, 55, 60, 61, 62, 64, 68, 69, 72, 73, 75, 76, 100, 112, 115, 116,
         120, 124, 137, 149, 152, 158, 159, 164},
        {2, 3, 9, 14, 17, 18, 20, 21, 22, 27, 30, 35, 36, 38, 40, 49, 75, 77, 87, 94, 95, 99, 100, 107, 109, 110, 115,
         116, 120, 131, 133, 149, 152, 155, 159},
        {20, 29, 30, 35, 38, 49, 80, 81, 86, 87, 88, 94, 95, 98, 99, 100, 102, 114, 123, 124, 125, 126, 131, 135, 136,
         137, 140, 141, 142, 149, 150, 152, 153},
        {2, 3, 6, 9, 10, 12, 14, 18, 20, 22, 25, 26, 27, 28, 37, 39, 75, 76, 99, 100, 103, 113, 115, 116, 120, 136, 137,
         140, 141, 148, 149, 150, 152, 153},
        {1, 3, 9, 10, 12, 13, 37, 41, 42, 43, 44, 45, 48, 50, 52, 54, 58, 87, 94, 95, 100, 120, 121, 125, 131, 136, 137,
         139, 140, 143, 144, 145, 149, 152},
        {2, 20, 22, 29, 30, 35, 36, 38, 49, 87, 94, 95, 101, 105, 106, 108, 115, 116, 120, 131, 135, 136, 137, 140, 141,
         142, 147, 148, 149, 150, 152, 153},
        {1, 7, 8, 11, 13, 26, 28, 29, 37, 39, 50, 53, 55, 59, 60, 61, 62, 68, 69, 72, 75, 76, 100, 112, 115, 116, 120,
         124, 136, 137, 148, 149, 152, 156},
        {2, 3, 5, 6, 9, 10, 12, 14, 18, 20, 22, 25, 26, 27, 28, 37, 39, 75, 76, 99, 100, 103, 112, 113, 115, 116, 120,
         136, 137, 140, 141, 148, 151, 154},
        {20, 25, 26, 27, 28, 31, 32, 37, 39, 75, 77, 80, 81, 84, 86, 87, 91, 94, 95, 98, 111, 114, 123, 124, 125, 126,
         131, 134, 136, 137, 140, 141, 142},
        {1, 3, 4, 9, 10, 12, 13, 15, 37, 41, 42, 43, 44, 45, 48, 87, 94, 95, 100, 120, 121, 125, 131, 136, 137, 139,
         140, 143, 144, 145, 149, 152, 158},
        {2, 3, 6, 9, 14, 17, 18, 19, 20, 21, 22, 27, 30, 31, 35, 36, 38, 40, 49, 75, 77, 87, 94, 95, 115, 116, 120, 131,
         132, 133, 149, 152, 155, 159},
        {2, 7, 8, 11, 14, 17, 20, 21, 22, 30, 35, 47, 49, 87, 94, 95, 100, 120, 125, 131, 136, 137, 139, 140, 143, 144,
         145, 149, 152, 158, 159, 160},
        {1, 7, 8, 11, 13, 28, 37, 39, 51, 52, 54, 57, 58, 75, 76, 99, 100, 112, 113, 115, 116, 120, 124, 136, 137, 140,
         141, 148, 149, 150, 152, 153},
        {27, 31, 35, 36, 49, 50, 53, 55, 60, 61, 62, 64, 68, 69, 72, 73, 74, 75, 76, 100, 112, 115, 116, 120, 124, 125,
         136, 137, 140, 141, 148, 152},
        {1, 7, 8, 13, 28, 29, 37, 39, 51, 52, 54, 56, 57, 58, 75, 76, 100, 112, 113, 115, 116, 120, 124, 136, 137, 140,
         141, 148, 149, 150, 152, 156},
        {3, 4, 5, 6, 9, 10, 12, 14, 18, 20, 26, 27, 28, 37, 39, 54, 58, 75, 76, 80, 81, 94, 96, 99, 100, 112, 113, 137,
         140, 141, 149, 150, 152, 153},
        {27, 31, 32, 35, 36, 40, 41, 42, 47, 49, 58, 63, 65, 66, 67, 70, 71, 74, 75, 76, 100, 112, 115, 116, 120, 124,
         137, 149, 152, 158, 159, 164},
        {25, 26, 28, 29, 33, 37, 38, 39, 50, 53, 55, 60, 61, 62, 64, 68, 69, 72, 75, 76, 100, 112, 115, 116, 120, 124,
         137, 149, 152, 158, 159, 164},
        {1, 2, 7, 8, 11, 13, 15, 23, 24, 26, 27, 28, 37, 39, 54, 58, 75, 76, 80, 81, 94, 96, 99, 100, 112, 113, 137,
         140, 141, 149, 150, 152, 153},
        {1, 4, 8, 9, 10, 11, 12, 13, 15, 27, 30, 35, 38, 49, 87, 94, 95, 99, 100, 104, 115, 116, 120, 131, 149, 152,
         156, 158, 159, 164, 165},
        {2, 3, 5, 6, 9, 10, 12, 14, 18, 19, 20, 22, 25, 26, 28, 37, 39, 78, 79, 101, 115, 116, 120, 131, 132, 136, 137,
         148, 149, 152, 155},
        {2, 4, 5, 6, 7, 8, 11, 14, 16, 17, 20, 21, 35, 49, 54, 58, 80, 81, 86, 87, 88, 100, 114, 124, 126, 131, 136,
         137, 140, 141, 142},
        {20, 25, 26, 27, 28, 32, 37, 39, 75, 77, 80, 82, 83, 86, 87, 89, 94, 95, 111, 115, 118, 119, 120, 121, 128, 131,
         134, 138, 146},
        {2, 7, 14, 16, 17, 18, 20, 21, 22, 30, 35, 38, 44, 46, 47, 49, 87, 94, 95, 120, 131, 139, 140, 149, 152, 156,
         158, 159, 162},
        {2, 5, 6, 7, 8, 11, 14, 16, 17, 20, 21, 22, 37, 41, 42, 43, 45, 48, 87, 94, 95, 100, 120, 125, 131, 139, 140,
         149, 152, 161},
        {20, 25, 26, 27, 28, 37, 39, 75, 77, 85, 94, 95, 111, 115, 118, 119, 120, 121, 122, 128, 129, 130, 131, 136,
         137, 140, 142},
        {25, 26, 27, 28, 31, 32, 37, 39, 40, 54, 58, 80, 81, 86, 87, 88, 97, 101, 105, 108, 137, 140, 141, 142, 149,
         150, 152, 153},
        {1, 3, 4, 9, 10, 12, 13, 15, 37, 41, 42, 43, 44, 45, 48, 87, 94, 95, 100, 120, 125, 131, 139, 140, 149, 152,
         158, 159, 163},
        {2, 6, 7, 8, 11, 14, 16, 17, 20, 21, 22, 35, 47, 49, 87, 94, 95, 100, 114, 122, 123, 124, 126, 127, 131, 158,
         159, 160},
        {1, 7, 8, 11, 13, 28, 29, 37, 39, 51, 52, 54, 56, 57, 58, 75, 76, 100, 112, 113, 115, 116, 120, 124, 137, 151,
         157},
        {20, 25, 26, 27, 28, 37, 39, 75, 77, 85, 93, 94, 95, 98, 99, 100, 107, 109, 110, 111, 131, 134, 137, 140, 141,
         142},
        {2, 20, 22, 29, 30, 34, 36, 38, 40, 49, 99, 100, 104, 112, 113, 114, 117, 136, 137, 140, 141, 148, 149, 150},
        {3, 4, 5, 6, 9, 10, 12, 14, 18, 19, 20, 26, 27, 28, 37, 39, 80, 81, 84, 86, 87, 90, 101, 114, 137, 140, 141}

    ]

    print("=" * 70)
    print("baseline PSO")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("Program completed")


if __name__ == "__main__":
    main()