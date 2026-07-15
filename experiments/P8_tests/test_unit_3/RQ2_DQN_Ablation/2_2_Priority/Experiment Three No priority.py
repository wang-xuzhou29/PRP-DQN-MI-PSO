import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# X, Y, Z - 150
X_MIN = 2
X_MAX = 100
Y_MIN = 2
Y_MAX = 100
Z_MIN = 2
Z_MAX = 100


# ===  ===
def normalize_state(state):
    """
    (1-50)(0-1)
    state: (x, y, z) tuple or array
    returns: normalized state (0-1 range)
    """
    x, y, z = state
    norm_x = (x - X_MIN) / (X_MAX - X_MIN)
    norm_y = (y - Y_MIN) / (Y_MAX - Y_MIN)
    norm_z = (z - Z_MIN) / (Z_MAX - Z_MIN)
    return (norm_x, norm_y, norm_z)


def denormalize_state(norm_state):
    """
    (0-1)(1-50)
    norm_state: normalized (x, y, z) tuple or array
    returns: denormalized state (1-50 range)
    """
    norm_x, norm_y, norm_z = norm_state
    x = int(round(norm_x * (X_MAX - X_MIN) + X_MIN))
    y = int(round(norm_y * (Y_MAX - Y_MIN) + Y_MIN))
    z = int(round(norm_z * (Z_MAX - Z_MIN) + Z_MIN))
    # 
    x = max(X_MIN, min(X_MAX, x))
    y = max(Y_MIN, min(Y_MAX, y))
    z = max(Z_MIN, min(Z_MAX, z))
    return (x, y, z)


def is_valid_state(state):
    """"""
    x, y, z = state
    return (X_MIN <= x <= X_MAX and
            Y_MIN <= y <= Y_MAX and
            Z_MIN <= z <= Z_MAX)


def clip_state(state):
    """"""
    x, y, z = state
    return (
        max(X_MIN, min(X_MAX, x)),
        max(Y_MIN, min(Y_MAX, y)),
        max(Z_MIN, min(Z_MAX, z))
    )


# === Metric() ===
class EnhancedStandardMetricsCollector:
    def __init__(self, experiment_name="Enhanced_Standard_DQN"):
        self.experiment_name = experiment_name
        self.start_time = None
        self.end_time = None

        # Metric
        self.total_reward = 0
        self.td_errors = []
        self.final_output_similarities = []
        self.action_improvements = []
        self.total_memory_usage = 0
        self.memory_check_count = 0
        self.step_count = 0

        # Metric()
        self.episode_rewards = []  # episode
        self.episode_similarities = []  # episodeAverage Similarity
        self.episode_td_errors = []  # episodeTD
        self.episode_epsilon_values = []  # epsilon
        self.episode_memory_usage = []  # episode

        # Path Metric
        self.similar_paths_performance = []
        self.isolated_paths_performance = []

        # ()
        self.milestone_data = {}  # episode 50, 100, 150, 200

        # 
        self.convergence_window = 20  # 
        self.convergence_threshold = 0.02  # 
        self.convergence_detected_episode = None

        # Metric
        self.sample_efficiency_data = []  # (episode, )
        self.performance_milestones = [0.6, 0.7, 0.75, 0.8]  # 

        # 
        self.learning_curve_characteristics = {}
        self.early_vs_late_performance = {}

        # X, Y, Z
        self.xyz_coordinate_stats = {
            'x_values': [],
            'y_values': [],
            'z_values': [],
            'x_distribution': {},
            'y_distribution': {},
            'z_distribution': {}
        }

    def reset(self):
        """Metric, """
        self.start_time = None
        self.end_time = None
        self.total_reward = 0
        self.td_errors = []
        self.final_output_similarities = []
        self.action_improvements = []
        self.total_memory_usage = 0
        self.memory_check_count = 0
        self.step_count = 0
        self.episode_rewards = []
        self.episode_similarities = []
        self.episode_td_errors = []
        self.episode_epsilon_values = []
        self.episode_memory_usage = []
        self.similar_paths_performance = []
        self.isolated_paths_performance = []
        self.milestone_data = {}
        self.convergence_detected_episode = None
        self.sample_efficiency_data = []
        self.learning_curve_characteristics = {}
        self.early_vs_late_performance = {}
        self.xyz_coordinate_stats = {
            'x_values': [],
            'y_values': [],
            'z_values': [],
            'x_distribution': {},
            'y_distribution': {},
            'z_distribution': {}
        }

    def start_training(self):
        self.start_time = time.time()

    def end_training(self):
        self.end_time = time.time()

    def record_xyz_coordinates(self, x_val, y_val, z_val):
        """X, Y, Z"""
        self.xyz_coordinate_stats['x_values'].append(x_val)
        self.xyz_coordinate_stats['y_values'].append(y_val)
        self.xyz_coordinate_stats['z_values'].append(z_val)

        self.xyz_coordinate_stats['x_distribution'][x_val] = self.xyz_coordinate_stats['x_distribution'].get(x_val,
                                                                                                             0) + 1
        self.xyz_coordinate_stats['y_distribution'][y_val] = self.xyz_coordinate_stats['y_distribution'].get(y_val,
                                                                                                             0) + 1
        self.xyz_coordinate_stats['z_distribution'][z_val] = self.xyz_coordinate_stats['z_distribution'].get(z_val,
                                                                                                             0) + 1

    def record_step_metrics(self, reward, td_error, triggered, target_path, x_coord=None, y_coord=None, z_coord=None):
        """Metric"""
        self.step_count += 1
        self.total_reward += reward
        self.td_errors.append(td_error)

        if x_coord is not None and y_coord is not None and z_coord is not None:
            self.record_xyz_coordinates(x_coord, y_coord, z_coord)

        process = psutil.Process(os.getpid())
        current_memory = process.memory_info().rss / 1024 / 1024
        self.total_memory_usage += current_memory
        self.memory_check_count += 1

    def record_episode_metrics(self, episode, episode_reward, avg_similarity, avg_td_error, epsilon,
                               path_group="similar"):
        """episodeMetric"""
        self.episode_rewards.append(episode_reward)
        self.episode_similarities.append(avg_similarity)
        self.episode_td_errors.append(avg_td_error)
        self.episode_epsilon_values.append(epsilon)

        process = psutil.Process(os.getpid())
        current_memory = process.memory_info().rss / 1024 / 1024
        self.episode_memory_usage.append(current_memory)

        if path_group == "similar":
            self.similar_paths_performance.append({
                'episode': episode,
                'reward': episode_reward,
                'similarity': avg_similarity,
                'td_error': avg_td_error
            })
        else:
            self.isolated_paths_performance.append({
                'episode': episode,
                'reward': episode_reward,
                'similarity': avg_similarity,
                'td_error': avg_td_error
            })

        if episode in [50, 100, 150, 200, 250, 300, 400, 450, 500]:
            self.milestone_data[episode] = {
                'avg_reward': np.mean(self.episode_rewards[-10:]) if len(
                    self.episode_rewards) >= 10 else episode_reward,
                'avg_similarity': avg_similarity,
                'avg_td_error': avg_td_error,
                'epsilon': epsilon,
                'memory_usage': current_memory,
                'total_steps': self.step_count
            }

        self._check_convergence(episode)
        self._check_performance_milestones(episode, avg_similarity)

    def _check_convergence(self, episode):
        """"""
        if len(self.episode_similarities) >= self.convergence_window and self.convergence_detected_episode is None:
            recent_similarities = self.episode_similarities[-self.convergence_window:]
            if np.std(recent_similarities) < self.convergence_threshold:
                self.convergence_detected_episode = episode

    def _check_performance_milestones(self, episode, similarity):
        """"""
        for milestone in self.performance_milestones:
            if similarity >= milestone and not any(data[1] == milestone for data in self.sample_efficiency_data):
                self.sample_efficiency_data.append((episode, milestone, self.step_count))

    def record_final_output_sample(self, triggered, target_path):
        """final samplesSimilarity"""
        if len(triggered | target_path) > 0:
            similarity = len(triggered & target_path) / len(triggered | target_path)
        else:
            similarity = 0.0
        self.final_output_similarities.append(similarity)

    def record_action_improvement(self, current_reward, prev_reward):
        """"""
        if prev_reward is not None:
            improvement = current_reward - prev_reward
            self.action_improvements.append(1 if improvement > 0 else 0)


# Metric
enhanced_standard_metrics = EnhancedStandardMetricsCollector("Enhanced_Standard_DQN_No_Priority")


# === reward function ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward


def execute_Tr(a):
    """
    替换原有的编排规则函数
    参数a: 包含3个元素的元组或数组，分别对应path_depth, file_count, access_level
    返回: 触发的规则编号集合
    """
    path_depth, file_count, access_level = float(a[0]), int(a[1]), float(a[2])

    # 使用正确的变量名
    x, y, z = path_depth, file_count, access_level

    triggered = set()
    # 创建一个字典来存储b数组的值，用于跟踪哪些规则被触发
    b = {}
    x_score = max(0, 100 - x * 1.67)

    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[0] = 1
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((z * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[1] = 2
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * x) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[2] = 3
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[3] = 4
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * 43) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[4] = 5
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((74 * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[5] = 6
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (47 + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[6] = 7
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (x + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[7] = 8
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (y + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[8] = 9
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 10) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[9] = 10
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) != 50 and x ** 2 + y ** 2 > z ** 2):
        b[10] = 11
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 29 and x ** 2 + y ** 2 > z ** 2):
        b[11] = 12
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 69 and x ** 2 + y ** 2 > z ** 2):
        b[12] = 13
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2):
        b[13] = 14
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and y ** 2 + y ** 2 > z ** 2):
        b[14] = 15
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 3 + y ** 2 > z ** 2):
        b[15] = 16
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + x ** 2 > z ** 2):
        b[16] = 17
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 3 > z ** 2):
        b[17] = 18
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 != z ** 2):
        b[18] = 19
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2.5):
        b[19] = 20
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 1.4 > z ** 2):
        b[20] = 21
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + 34 ** 2 > z ** 2):
        b[21] = 22
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and 64 ** 2 + y ** 2 > z ** 2):
        b[22] = 23
    if (x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2:
        pattern_type = "311A1"
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((z ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[23] = 24
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - z ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[24] = 25
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 1.5) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[25] = 26
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2.3 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[26] = 27
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((45 ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[27] = 28
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - 31 ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[28] = 29
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[29] = 30
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (x + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[30] = 31
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 2) < -30 or (abs(x - y) * z) / 100 > 45):
        b[31] = 32
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (26 + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[32] = 33
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y - 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[33] = 34
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) != -30 or (abs(x - y) * z) / 100 > 45):
        b[34] = 35
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -10 or (abs(x - y) * z) / 100 > 45):
        b[35] = 36
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 and (abs(x - y) * z) / 100 > 45):
        b[36] = 37
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(z - y) * z) / 100 > 45):
        b[37] = 38
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - z) * z) / 100 > 45):
        b[38] = 39
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * y) / 100 > 45):
        b[39] = 40
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * x) / 100 > 45):
        b[40] = 41
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * 34) / 100 > 45):
        b[41] = 42
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - 26) * z) / 100 > 45):
        b[42] = 43
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(52 - y) * z) / 100 > 45):
        b[43] = 44
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 76 > 45):
        b[44] = 45
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 != 45):
        b[45] = 46
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 55):
        b[46] = 47
    if (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45:
        pattern_type = "311A2"

    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + x) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35):
        b[47] = 48
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((y + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35):
        b[48] = 49
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((24 + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35):
        b[49] = 50
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + 31) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35):
        b[50] = 51
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 1.5) ** 2 < z * 20 and (x * x * z) / 1000 < 35):
        b[51] = 52
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2.4 < z * 20 and (x * x * z) / 1000 < 35):
        b[52] = 53
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 != z * 20 and (x * x * z) / 1000 < 35):
        b[53] = 54
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < 25 * 20 and (x * x * z) / 1000 < 35):
        b[54] = 55
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 14 and (x * x * z) / 1000 < 35):
        b[55] = 56
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < x * 20 and (x * x * z) / 1000 < 35):
        b[56] = 57
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < y * 20 and (x * x * z) / 1000 < 35):
        b[57] = 58
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 1.2 < z * 20 and (x * x * z) / 1000 < 35):
        b[58] = 59
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 or (x * x * z) / 1000 < 35):
        b[59] = 60
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (y * x * z) / 1000 < 35):
        b[60] = 61
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (z * x * z) / 1000 < 35):
        b[61] = 62
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * 35 * z) / 1000 < 35):
        b[62] = 63
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * x * x) / 1000 < 35):
        b[63] = 64
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * x * y) / 1000 < 35):
        b[64] = 65
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * x * 26) / 1000 < 35):
        b[65] = 66
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * 54 * z) / 1000 < 35):
        b[66] = 67
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1050 < 35):
        b[67] = 68
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 != 35):
        b[68] = 69
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 25):
        b[69] = 70
    if ((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35:
        pattern_type = "311A3"
    if (x < 35 and y < 50) != (x < 35 or y < 50):
        b[70] = 71
    if (x < 35 and y < 50) != (x != 35 and y < 50):
        b[71] = 72
    if (x < 35 and y < 50) != (x < 19 and y < 50):
        b[72] = 73
    if (x < 35 and y < 50) != (x < 35 and y != 50):
        b[73] = 74
    if (x < 35 and y < 50) != (x < 35 and y < 61):
        b[74] = 75
    if x < 35 and y < 50:
        pattern_type = 312
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y < 35 or z < 65 and x < 50)):
        b[75] = 76
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y != 35 and z < 65 and x < 50)):
        b[76] = 77
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y < 25 and z < 65 and x < 50)):
        b[77] = 78
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y < 35 and z != 65 and x < 50)):
        b[78] = 79
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y < 35 and z < 72 and x < 50)):
        b[79] = 80
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y < 35 and z < 65 or x < 50)):
        b[80] = 81
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y < 35 and z < 65 and x != 50)):
        b[81] = 82
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 31)):
        b[82] = 83
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y < 35 and z < 53 and x < 50)):
        b[83] = 84
    if (y < 30 and z < 60) or (y < 35 and z < 65 and x < 50):
        pattern_type = 313
    if (z < 40 and x > 50) != (z < 40 or x > 50):
        b[84] = 85
    if (z < 40 and x > 50) != (z < 40 and x != 50):
        b[85] = 86
    if (z < 40 and x > 50) != (z != 40 and x > 50):
        b[86] = 87
    if (z < 40 and x > 50) != (z < 29 and x > 50):
        b[87] = 88
    if (z < 40 and x > 50) != (z < 40 and x > 62):
        b[88] = 89
    if z < 40 and x > 50:
        pattern_type = 314
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (y * y) / 100 < 28) or (x < 60 and y < 55)):
        b[89] = 90
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * x) / 100 < 28) or (x < 60 and y < 55)):
        b[90] = 91
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * 53) / 100 < 28) or (x < 60 and y < 55)):
        b[91] = 92
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (41 * y) / 100 < 28) or (x < 60 and y < 55)):
        b[92] = 93
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 88 < 28) or (x < 60 and y < 55)):
        b[93] = 94
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 100 != 28) or (x < 60 and y < 55)):
        b[94] = 95
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 100 < 21) or (x < 60 and y < 55)):
        b[95] = 96
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) and (x < 60 and y < 55)):
        b[96] = 97
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 or y < 55)):
        b[97] = 98
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x != 60 and y < 55)):
        b[98] = 99
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 69 and y < 55)):
        b[99] = 100
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y != 55)):
        b[100] = 101
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 45)):
        b[101] = 102
    if (40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55):
        pattern_type = 315
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((x ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)):
        b[102] = 103
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((z ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)):
        b[103] = 104
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2.4) / (x + z + 1) < 22) or (y < 65 and x < 70)):
        b[104] = 105
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (z + z + 1) < 22) or (y < 65 and x < 70)):
        b[105] = 106
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (y + z + 1) < 22) or (y < 65 and x < 70)):
        b[106] = 107
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + x + 1) < 22) or (y < 65 and x < 70)):
        b[107] = 108
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + y + 1) < 22) or (y < 65 and x < 70)):
        b[108] = 109
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 7) < 22) or (y < 65 and x < 70)):
        b[109] = 110
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 1) != 22) or (y < 65 and x < 70)):
        b[110] = 111
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 1) < 32) or (y < 65 and x < 70)):
        b[111] = 112
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 1) < 22) and (y < 65 and x < 70)):
        b[112] = 113
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 1) < 22) or (y != 65 and x < 70)):
        b[113] = 114
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 1) < 22) or (y < 45 and x < 70)):
        b[114] = 115
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 1) < 22) or (y < 65 or x < 70)):
        b[115] = 116
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x != 70)):
        b[116] = 117
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 55)):
        b[117] = 118
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((24 ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)):
        b[118] = 119
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (56 + z + 1) < 22) or (y < 65 and x < 70)):
        b[119] = 120
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + 73 + 1) < 22) or (y < 65 and x < 70)):
        b[120] = 121
    if ((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70):
        pattern_type = 316
    if (50 <= y < 75 and x > 65) != (50 <= y < 75 or x > 65):
        b[121] = 122
    if (50 <= y < 75 and x > 65) != (50 <= y < 75 and x != 65):
        b[122] = 123
    if (50 <= y < 75 and x > 65) != (50 <= y < 75 and x > 47):
        b[123] = 124
    if (50 <= y < 75 and x > 65) != (50 <= y < 82 and x > 65):
        b[124] = 125
    if (50 <= y < 75 and x > 65) != (57 <= y < 75 and x > 65):
        b[125] = 126
    if 50 <= y < 75 and x > 65:
        pattern_type = 317
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 89 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)):
        b[126] = 127
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((41 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)):
        b[127] = 128
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 or y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)):
        b[128] = 129
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y != 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)):
        b[129] = 130
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 79 and (z * y) / 100 < 52) or (z < 75 and y > 75)):
        b[130] = 131
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 or (z * y) / 100 < 52) or (z < 75 and y > 75)):
        b[131] = 132
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (y * y) / 100 < 52) or (z < 75 and y > 75)):
        b[132] = 133
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * z) / 100 < 52) or (z < 75 and y > 75)):
        b[133] = 134
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * 26) / 100 < 52) or (z < 75 and y > 75)):
        b[134] = 135
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (53 * y) / 100 < 52) or (z < 75 and y > 75)):
        b[135] = 136
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * y) / 89 < 52) or (z < 75 and y > 75)):
        b[136] = 137
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * y) / 100 != 52) or (z < 75 and y > 75)):
        b[137] = 138
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * y) / 100 < 39) or (z < 75 and y > 75)):
        b[138] = 139
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) and (z < 75 and y > 75)):
        b[139] = 140
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z != 75 and y > 75)):
        b[140] = 141
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 65 and y > 75)):
        b[141] = 142
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 or y > 75)):
        b[142] = 143
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y != 75)):
        b[143] = 144
    if (55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75):
        pattern_type = 318
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x != 88 and y > 75) or (78 <= x < 90 and y > 72)):
        b[144] = 145
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x < 94 and y > 75) or (78 <= x < 90 and y > 72)):
        b[145] = 146
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x < 88 or y > 75) or (78 <= x < 90 and y > 72)):
        b[146] = 147
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x < 88 and y != 75) or (78 <= x < 90 and y > 72)):
        b[147] = 148
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x < 88 and y > 67) or (78 <= x < 90 and y > 72)):
        b[148] = 149
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x < 88 and y > 75) and (78 <= x < 90 and y > 72)):
        b[149] = 150
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x < 88 and y > 75) or (78 <= x < 90 or y > 72)):
        b[150] = 151
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y != 72)):
        b[151] = 152
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 62)):
        b[152] = 153
    if (75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72):
        pattern_type = 319
    if (x >= 88 and y >= 85 and z >= 85) != (x >= 88 or y >= 85 and z >= 85):
        b[153] = 154
    if (x >= 88 and y >= 85 and z >= 85) != (x != 88 and y >= 85 and z >= 85):
        b[154] = 155
    if (x >= 88 and y >= 85 and z >= 85) != (x >= 78 and y >= 85 and z >= 85):
        b[155] = 156
    if (x >= 88 and y >= 85 and z >= 85) != (x >= 88 and y >= 85 or z >= 85):
        b[156] = 157
    if (x >= 88 and y >= 85 and z >= 85) != (x >= 88 and y != 85 and z >= 85):
        b[157] = 158
    if (x >= 88 and y >= 85 and z >= 85) != (x >= 88 and y >= 85 and z != 85):
        b[158] = 159
    if x >= 88 and y >= 85 and z >= 85:
        pattern_type = 320

    # 返回被触发的规则编号集合
    return set(b.values())
def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0


# === Path Similarity ===
def compute_path_similarity_matrix(paths):
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix


targetPaths = [
    {1, 6, 11, 14, 25, 26, 29, 37, 48, 51, 54, 58, 59, 60, 71, 74, 76, 77, 81, 85, 86, 98, 101, 103, 111, 114, 116, 119,
     122, 123, 129, 132, 143, 144, 151},
    {1, 4, 5, 6, 11, 14, 25, 26, 29, 37, 54, 59, 60, 71, 74, 76, 77, 81, 85, 86, 98, 101, 103, 104, 111, 114, 116, 119,
     122, 123, 129, 132, 143, 144, 151},
    {1, 2, 6, 8, 11, 14, 24, 26, 29, 37, 48, 51, 54, 59, 60, 71, 74, 76, 77, 81, 98, 101, 103, 111, 114, 116, 119, 122,
     123, 131, 133, 137, 139, 140, 151},
    {1, 2, 7, 8, 9, 10, 13, 24, 35, 46, 54, 60, 71, 72, 98, 99, 104, 108, 109, 110, 111, 112, 116, 117, 119, 121, 122,
     132, 143, 144, 147, 148, 151, 152},
    {1, 3, 4, 5, 6, 11, 14, 25, 26, 27, 29, 37, 54, 60, 71, 74, 76, 77, 81, 85, 86, 98, 101, 104, 111, 114, 116, 119,
     122, 123, 129, 132, 143, 144, 151},
    {1, 2, 6, 8, 11, 14, 24, 26, 29, 37, 48, 51, 54, 59, 60, 71, 74, 81, 98, 101, 103, 111, 114, 116, 119, 129, 132,
     135, 136, 138, 141, 143, 147, 151},
    {1, 2, 6, 11, 14, 49, 50, 52, 53, 55, 56, 57, 71, 74, 81, 98, 101, 103, 111, 114, 116, 119, 122, 123, 129, 132, 134,
     135, 136, 138, 143, 144, 151},
    {1, 2, 6, 11, 14, 24, 26, 29, 37, 48, 51, 54, 58, 59, 60, 71, 74, 76, 77, 81, 98, 101, 103, 111, 114, 116, 119, 122,
     123, 128, 129, 132, 143, 144},
    {1, 2, 6, 11, 14, 24, 26, 29, 37, 49, 50, 52, 53, 55, 56, 57, 71, 74, 76, 77, 81, 98, 101, 103, 111, 116, 119, 122,
     123, 129, 130, 132, 143, 144},
    {1, 3, 4, 5, 6, 11, 12, 14, 24, 28, 30, 31, 33, 34, 35, 36, 46, 54, 60, 89, 98, 101, 104, 111, 114, 116, 119, 122,
     123, 124, 129, 132, 143, 144},
    {1, 4, 5, 15, 20, 23, 37, 38, 39, 40, 42, 44, 47, 71, 72, 76, 85, 87, 98, 99, 103, 104, 113, 122, 129, 130, 132,
     143, 144, 147, 148, 151, 152},
    {3, 5, 9, 10, 13, 17, 20, 21, 22, 24, 26, 27, 29, 37, 85, 87, 98, 101, 103, 111, 114, 116, 119, 129, 132, 135, 136,
     138, 141, 143, 147, 151},
    {1, 4, 15, 20, 23, 37, 38, 39, 40, 42, 43, 44, 71, 72, 76, 85, 87, 98, 99, 103, 104, 113, 122, 129, 130, 132, 143,
     144, 147, 148, 151, 152},
    {1, 2, 6, 11, 14, 24, 26, 28, 29, 37, 49, 50, 52, 53, 55, 56, 57, 71, 74, 75, 76, 77, 81, 98, 113, 115, 122, 123,
     129, 130, 132, 143, 144},
    {24, 26, 28, 29, 37, 49, 50, 52, 53, 55, 56, 57, 58, 61, 62, 63, 67, 71, 74, 75, 81, 97, 102, 113, 115, 122, 123,
     129, 130, 132, 143, 144},
    {24, 26, 28, 29, 37, 49, 50, 52, 53, 55, 56, 57, 58, 61, 62, 63, 67, 70, 71, 74, 75, 81, 98, 113, 115, 122, 123,
     129, 130, 132, 143, 144},
    {1, 2, 6, 7, 8, 9, 10, 13, 24, 35, 46, 54, 60, 71, 72, 98, 99, 104, 108, 109, 110, 111, 112, 116, 117, 119, 121,
     122, 132, 143, 144, 154},
    {1, 2, 7, 8, 9, 10, 13, 24, 35, 46, 54, 60, 71, 72, 76, 81, 82, 98, 99, 103, 105, 106, 107, 113, 122, 132, 143, 144,
     147, 148, 151, 152},
    {1, 4, 15, 20, 35, 41, 45, 46, 71, 72, 76, 81, 82, 85, 87, 98, 99, 103, 104, 105, 113, 122, 129, 130, 132, 143, 144,
     147, 148, 151, 152},
    {1, 2, 7, 8, 9, 10, 13, 24, 35, 46, 54, 60, 98, 99, 104, 108, 109, 111, 112, 116, 117, 119, 121, 126, 132, 143, 144,
     147, 148, 151, 152},
    {14, 16, 17, 18, 19, 22, 37, 39, 40, 42, 43, 44, 47, 71, 72, 76, 85, 87, 98, 99, 103, 104, 113, 122, 129, 130, 132,
     147, 148, 151, 152},
    {1, 3, 4, 5, 6, 11, 12, 14, 24, 30, 33, 35, 36, 46, 54, 59, 60, 76, 77, 81, 85, 86, 90, 94, 96, 97, 113, 115, 122,
     123, 132, 143, 144},
    {1, 4, 5, 15, 20, 21, 22, 23, 25, 35, 38, 43, 45, 46, 71, 72, 85, 87, 98, 99, 103, 104, 105, 113, 122, 132, 151,
     152, 154, 157, 158},
    {3, 5, 13, 20, 21, 22, 24, 26, 27, 29, 30, 32, 37, 85, 87, 98, 101, 111, 114, 116, 119, 129, 132, 135, 136, 138,
     141, 143, 147, 151},
    {24, 26, 27, 28, 29, 37, 49, 50, 52, 53, 55, 56, 57, 58, 61, 62, 63, 67, 73, 76, 77, 81, 97, 102, 113, 115, 129,
     130, 132, 143, 144},
    {14, 16, 17, 19, 37, 39, 40, 42, 44, 60, 61, 65, 69, 71, 72, 76, 85, 87, 98, 99, 103, 104, 113, 122, 132, 147, 148,
     151, 152, 157},
    {2, 7, 8, 9, 10, 13, 24, 28, 30, 31, 33, 35, 36, 46, 88, 89, 98, 101, 104, 111, 114, 116, 119, 122, 123, 124, 129,
     132, 143, 144},
    {3, 5, 9, 13, 17, 20, 21, 22, 24, 26, 27, 29, 37, 85, 87, 98, 101, 103, 111, 114, 116, 119, 141, 143, 147, 151, 154,
     155, 157},
    {3, 4, 5, 6, 11, 12, 14, 24, 30, 35, 36, 46, 54, 59, 60, 89, 91, 92, 93, 95, 98, 101, 113, 115, 122, 123, 124, 132,
     143, 144},
    {1, 2, 6, 7, 8, 9, 10, 13, 24, 35, 46, 54, 60, 71, 72, 98, 99, 103, 105, 106, 107, 113, 120, 122, 132, 143, 144,
     147, 148},
    {1, 2, 6, 8, 11, 12, 14, 24, 26, 29, 37, 71, 74, 81, 98, 101, 103, 111, 114, 116, 119, 122, 123, 127, 131, 137, 139,
     140},
    {1, 3, 4, 5, 6, 12, 14, 24, 28, 30, 33, 35, 36, 46, 54, 59, 60, 89, 98, 101, 113, 115, 118, 122, 123, 124, 132, 143,
     144},
    {2, 6, 7, 8, 9, 10, 13, 24, 28, 30, 33, 35, 36, 46, 54, 60, 104, 111, 119, 122, 140, 145, 146, 147, 151, 154, 157,
     159},
    {3, 4, 5, 6, 11, 12, 14, 24, 35, 46, 54, 60, 89, 93, 95, 98, 99, 100, 113, 115, 118, 122, 123, 124, 132, 143, 144},
    {5, 20, 21, 22, 23, 28, 35, 43, 46, 85, 87, 111, 112, 119, 122, 125, 141, 143, 145, 146, 147, 151, 154, 157, 158},
    {5, 20, 28, 35, 46, 85, 87, 111, 112, 119, 129, 132, 133, 135, 136, 138, 143, 144, 147, 148, 149, 151, 152, 153},
    {5, 20, 24, 28, 33, 35, 36, 43, 46, 85, 87, 111, 119, 122, 129, 132, 135, 138, 141, 143, 150, 154, 157, 159},
    {1, 2, 4, 6, 7, 8, 11, 12, 14, 25, 26, 27, 29, 37, 81, 98, 101, 103, 111, 114, 116, 119, 140, 142, 147, 151},
    {24, 26, 27, 28, 29, 30, 37, 60, 64, 65, 66, 69, 71, 75, 81, 97, 102, 113, 115, 122, 123, 129, 130, 132},
    {6, 11, 14, 24, 25, 26, 27, 28, 29, 32, 37, 48, 54, 55, 58, 59, 60, 78, 85, 86, 97, 113, 132, 143, 144},
    {25, 31, 35, 36, 46, 53, 55, 57, 58, 61, 62, 63, 67, 70, 73, 76, 79, 81, 97, 129, 130, 132, 143, 144},
    {25, 35, 36, 46, 52, 53, 55, 56, 57, 58, 61, 62, 63, 67, 70, 73, 78, 84, 97, 129, 130, 132, 143, 144},
    {25, 35, 36, 46, 53, 55, 57, 58, 61, 62, 63, 67, 70, 73, 76, 79, 80, 81, 97, 129, 130, 132, 143, 144},
    {5, 20, 21, 22, 28, 30, 31, 33, 35, 36, 43, 46, 85, 87, 111, 119, 122, 141, 143, 154, 155, 156, 157},
    {24, 26, 28, 29, 30, 37, 60, 64, 65, 66, 68, 69, 71, 74, 75, 81, 98, 101, 113, 115, 122, 123, 157},
    {3, 4, 5, 6, 11, 12, 14, 24, 35, 46, 54, 59, 60, 71, 72, 78, 83, 85, 86, 97, 113, 132, 143, 144},

]


# === Path  ===
def group_paths_by_similarity(paths):
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.mean(avg_sim_scores)

    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]
    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)

    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group


# === (, )===
class StandardExperienceReplay:
    def __init__(self, capacity=10000):
        self.capacity = capacity
        self.buffer = deque(maxlen=self.capacity)

    def append(self, experience):
        """"""
        self.buffer.append(experience[:5])

    def sample(self, batch_size):
        """
        
        , 
        """
        if len(self.buffer) < batch_size:
            return [], [], None
        # replace=False 
        batch_indices = np.random.choice(len(self.buffer), batch_size, replace=False)
        batch = [self.buffer[idx] for idx in batch_indices]
        return batch, batch_indices, None

    def __len__(self):
        return len(self.buffer)

    def get_high_reward_samples(self, target_path, num_samples=20):
        """, """
        if len(self.buffer) == 0:
            return []
        samples_with_recalculated_scores = []
        for experience in self.buffer:
            # , 
            norm_state_tensor = experience[0]
            norm_state = tuple(norm_state_tensor.cpu().numpy().flatten())
            # 
            state_tuple = denormalize_state(norm_state)
            triggered = execute_Tr(state_tuple[0], state_tuple[1], state_tuple[2])
            new_reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_recalculated_scores.append((state_tuple, new_reward, sim, triggered))
        samples_with_recalculated_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_recalculated_scores[:num_samples]


def load_path_data(file_path):
    path_data = []
    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines[2:]:
            parts = line.strip().split("\t")
            state = tuple(map(int, parts[0].split()))
            path_data.append(state)
    return path_data


# === DQN ===
class DQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, 128)
        self.fc2 = nn.Linear(128, 64)
        self.fc3 = nn.Linear(64, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)


# === DQN Agent()===
class StandardDQNAgent:
    def __init__(self, state_dim, action_dim, replay_buffer, gamma=0.99, epsilon=1.0,
                 epsilon_decay=0.995, epsilon_min=0.1, learning_rate=0.001):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        delta_values = [10, 5, 3, 2, 1, -1, -2, -3, -5, -10]
        dim = action_idx // 10
        delta_idx = action_idx % 10
        delta = delta_values[delta_idx]
        if dim == 0:
            return (delta, 0, 0)
        elif dim == 1:
            return (0, delta, 0)
        elif dim == 2:
            return (0, 0, delta)

    def act(self, state):
        """
        
        :  (1-50)
        """
        if random.random() < self.epsilon:
            return random.randrange(self.action_dim)
        # 
        norm_state = normalize_state(state)
        state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)
        return torch.argmax(q_values, dim=1).item()

    def store_transition(self, state, action, reward, next_state, done):
        """
        , 
        :  (1-50)
        :  (0-1)
        """
        # 
        norm_state = normalize_state(state)
        norm_next_state = normalize_state(next_state)

        # tensor
        norm_state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
        norm_next_state_tensor = torch.tensor(norm_next_state, dtype=torch.float32).unsqueeze(0).to(device)

        # TD
        with torch.no_grad():
            q_values = self.model(norm_state_tensor)
            next_q_values = self.target_model(norm_next_state_tensor)
            max_next_q_values = next_q_values.max(1)[0]
            target_q_values = reward + (self.gamma * max_next_q_values * (1 - done))
            td_error = torch.abs(q_values[0][action] - target_q_values).item()

        # 
        self.replay_buffer.append((norm_state_tensor, action, reward, norm_next_state_tensor, done))
        return td_error

    def train(self, batch_size=32):
        """
        
        , 
        """
        if len(self.replay_buffer) < batch_size:
            return

        # ()
        batch, batch_indices, _ = self.replay_buffer.sample(batch_size)
        states, actions, rewards, next_states, dones = zip(*batch)

        # tensor(, )
        states = torch.tensor(np.array([s.cpu().numpy().flatten() for s in states]), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array([ns.cpu().numpy().flatten() for ns in next_states]),
                                   dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        # Q
        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))

        loss = nn.MSELoss()(current_q_values, target_q_values)

        self.optimizer.zero_grad()
        loss.backward()
        self.optimizer.step()

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())


# === Sample generation===
def generate_samples_for_similar_paths(similar_group_indices, num_total=2000, top_k=200):
    def jaccard_similarity_local(a, b):
        if not a and not b: return 1.0
        return len(a & b) / len(a | b) if a | b else 0.0

    def compute_robustness(state, path):
        base = execute_Tr(state[0], state[1], state[2])
        if not base: return 0.0
        rob, neighbors = 0.0, 0
        for dx in [-1, 0, 1]:
            for dy in [-1, 0, 1]:
                for dz in [-1, 0, 1]:
                    if dx == dy == dz == 0: continue
                    neighbor_state = (state[0] + dx, state[1] + dy, state[2] + dz)
                    if not is_valid_state(neighbor_state): continue
                    neighbor = clip_state(neighbor_state)
                    n_trig = execute_Tr(neighbor[0], neighbor[1], neighbor[2])
                    if not n_trig: continue
                    rob += jaccard_similarity_local(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"enhanced_standard_path{path_id}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Enhanced Standard Path {path_id}\n")
            f.write("x y z\tScore\tSimilarity\tLengthDiff\tRobustness\n")
            for s in samples:
                x, y, z = s[0]
                f.write(f"{x} {y} {z}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_enhanced_standard"
    for path_idx in similar_group_indices:
        path = targetPaths[path_idx]
        samples = []
        attempts = 0
        while len(samples) < top_k and attempts < num_total * 5:
            attempts += 1
            state = (
                np.random.randint(X_MIN, X_MAX + 1),
                np.random.randint(Y_MIN, Y_MAX + 1),
                np.random.randint(Z_MIN, Z_MAX + 1)
            )
            triggered = execute_Tr(state[0], state[1], state[2])
            if not triggered: continue
            sim = jaccard_similarity_local(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            score = 0.55 * sim + 0.25 * len_diff + 0.2 * rob
            samples.append((state, score, sim, len_diff, rob))
        if samples:
            samples.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples[:top_k], base_dir=base_dir)


# === Run ===
def enhanced_standard_generate_and_train_for_similar_paths(agent, similar_group, path_documents, episodes=500,
                                                           batch_size=32, steps_per_test=5, replay_times=10,
                                                           is_isolated=False):
    trained_paths = set()
    update_target_every = 100
    global_steps = 0

    for episode in range(episodes):
        episode_reward = 0
        episode_similarities = []
        episode_td_errors = []

        for path_idx in similar_group:
            if path_idx in trained_paths:
                continue

            file_path = os.path.join(path_documents,
                                     f"enhanced_standard_path{path_idx + 1}{'_isolated' if is_isolated else ''}.txt")
            path_data = load_path_data(file_path)
            target_path = targetPaths[path_idx]

            BATCH_SIZE = 50
            N_SAMPLES = 200
            N_STEPS = 10
            REPLAY_TIMES = 3

            for batch_start in range(0, N_SAMPLES, BATCH_SIZE):
                batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)
                for test_data in range(batch_start, batch_end):
                    if test_data >= len(path_data):
                        break

                    step_count = 0
                    state = path_data[test_data]  #  (1-50)
                    prev_state = None
                    prev_triggered = None
                    prev_reward = None

                    for step in range(N_STEPS):
                        legal_actions = []
                        for a in range(agent.action_dim):
                            dx, dy, dz = agent.decode_action(a)
                            cand_next = (state[0] + dx, state[1] + dy, state[2] + dz)
                            if is_valid_state(cand_next):
                                legal_actions.append(a)
                        if not legal_actions:
                            break

                        if random.random() < agent.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            # act
                            norm_state = normalize_state(state)
                            state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        dx, dy, dz = agent.decode_action(action)
                        next_state = clip_state((state[0] + dx, state[1] + dy, state[2] + dz))

                        triggered = execute_Tr(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                        done = (step_count == N_STEPS - 1)

                        # store_transition
                        td_error = agent.store_transition(state, action, reward, next_state, done)
                        enhanced_standard_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                      next_state[0], next_state[1], next_state[2])

                        episode_similarities.append(jaccard_similarity(triggered, target_path))
                        episode_td_errors.append(td_error)

                        if prev_reward is not None:
                            enhanced_standard_metrics.record_action_improvement(reward, prev_reward)

                        prev_state = state
                        prev_triggered = triggered
                        prev_reward = reward
                        state = next_state
                        step_count += 1
                        episode_reward += reward
                        global_steps += 1

                        if global_steps % update_target_every == 0:
                            agent.update_target_model()

                for _ in range(REPLAY_TIMES):
                    if len(agent.replay_buffer) >= batch_size:
                        agent.train(batch_size)

            trained_paths.add(path_idx)

        avg_similarity = np.mean(episode_similarities) if episode_similarities else 0
        avg_td_error = np.mean(episode_td_errors) if episode_td_errors else 0
        enhanced_standard_metrics.record_episode_metrics(episode, episode_reward, avg_similarity, avg_td_error,
                                                         agent.epsilon, "similar")

        if episode % 10 == 0:
            agent.update_target_model()

        if len(trained_paths) == len(similar_group):
            break

    return agent


def generate_samples_for_isolated_paths_standard(agent_similar, isolated_group_indices, num_total=2000, top_k=200):
    def compute_q_value(state, agent):
        # Q
        norm_state = normalize_state(state)
        state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = agent.model(state_tensor)
        return q_values.max().item()

    def compute_robustness(state, path):
        base = execute_Tr(state[0], state[1], state[2])
        if not base: return 0.0
        rob, neighbors = 0.0, 0
        for dx in [-1, 0, 1]:
            for dy in [-1, 0, 1]:
                for dz in [-1, 0, 1]:
                    if dx == dy == dz == 0: continue
                    neighbor_state = (state[0] + dx, state[1] + dy, state[2] + dz)
                    if not is_valid_state(neighbor_state): continue
                    neighbor = clip_state(neighbor_state)
                    n_trig = execute_Tr(neighbor[0], neighbor[1], neighbor[2])
                    if not n_trig: continue
                    rob += jaccard_similarity(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"enhanced_standard_path{path_id}_isolated.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Enhanced Standard Isolated Path {path_id}\n")
            f.write("x y z\tScore\tSimilarity\tLengthDiff\tRobustness\tQ_complement\n")
            for s in samples:
                x, y, z = s[0]
                f.write(f"{x} {y} {z}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\t{s[5]:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_enhanced_standard"
    for path_idx in isolated_group_indices:
        path = targetPaths[path_idx]
        samples_raw = []  # Q
        attempts = 0

        # Run : candidatesQ
        print(f"Path  {path_idx + 1} ...")
        while len(samples_raw) < num_total and attempts < num_total * 5:
            attempts += 1
            state = (
                np.random.randint(X_MIN, X_MAX + 1),
                np.random.randint(Y_MIN, Y_MAX + 1),
                np.random.randint(Z_MIN, Z_MAX + 1)
            )
            triggered = execute_Tr(state[0], state[1], state[2])
            if not triggered: continue
            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            q_value = compute_q_value(state, agent_similar)
            samples_raw.append((state, sim, len_diff, rob, q_value))

        # Run : Q
        if not samples_raw:
            print(f"Path  {path_idx + 1}: ")
            continue

        q_values_list = [s[4] for s in samples_raw]
        q_min = min(q_values_list)
        q_max = max(q_values_list)

        print(f"Path  {path_idx + 1}:  {len(samples_raw)} ")
        print(f"  Q: [{q_min:.4f}, {q_max:.4f}]")

        # Run : Q, 
        samples_final = []
        for state, sim, len_diff, rob, q_value in samples_raw:
            # : (q - q_min) / (q_max - q_min)
            if q_max - q_min > 1e-6:  # 
                q_normalized = (q_value - q_min) / (q_max - q_min)
            else:
                q_normalized = 0.5  # Q, 0.5

            # : 1 - q_normalized
            q_complement = 1.0 - q_normalized

            # final samples, QRun Metric
            score = 0.28 * sim + 0.1 * len_diff + 0.19 * rob + 0.43 * q_complement

            samples_final.append((state, score, sim, len_diff, rob, q_complement))

        # Run : top_k
        if samples_final:
            samples_final.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples_final[:top_k], base_dir=base_dir)
            print(f"Path  {path_idx + 1}:  {min(top_k, len(samples_final))} ")
            print(f"  Q_complement : [{samples_final[-1][5]:.4f}, {samples_final[0][5]:.4f}]")


# === Run ===
def enhanced_standard_generate_and_train_for_isolated_paths(agent_similar, agent_isolated, similar_group,
                                                            isolated_group, path_documents, episodes=500, batch_size=32,
                                                            is_isolated=True):
    trained_paths = set()
    update_target_every = 100
    global_steps = 0

    stage1_samples_pool = {}

    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        # get_high_reward_samples
        high_reward_samples = agent_similar.replay_buffer.get_high_reward_samples(target_path, num_samples=100)
        stage1_samples_pool[path_idx] = high_reward_samples

    for episode in range(episodes):
        episode_reward = 0
        episode_similarities = []
        episode_td_errors = []

        for path_idx in isolated_group:
            if path_idx in trained_paths:
                continue

            file_path = os.path.join(path_documents, f"enhanced_standard_path{path_idx + 1}_isolated.txt")
            stage2_path_data = load_path_data(file_path)
            stage1_samples = stage1_samples_pool.get(path_idx, [])
            target_path = targetPaths[path_idx]

            BATCH_SIZE = 50
            N_SAMPLES_STAGE2 = min(140, len(stage2_path_data))
            N_SAMPLES_STAGE1 = min(60, len(stage1_samples))
            N_STEPS = 10
            REPLAY_TIMES = 3

            for batch_start in range(0, N_SAMPLES_STAGE2, BATCH_SIZE):
                batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES_STAGE2)

                for test_data_idx in range(batch_start, batch_end):
                    if test_data_idx >= len(stage2_path_data):
                        break

                    step_count = 0
                    state = stage2_path_data[test_data_idx]  # 
                    prev_state = None
                    prev_triggered = None
                    prev_reward = None

                    for step in range(N_STEPS):
                        legal_actions = []
                        for a in range(agent_isolated.action_dim):
                            dx, dy, dz = agent_isolated.decode_action(a)
                            cand_next = (state[0] + dx, state[1] + dy, state[2] + dz)
                            if is_valid_state(cand_next):
                                legal_actions.append(a)

                        if not legal_actions:
                            break

                        if random.random() < agent_isolated.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            # 
                            norm_state = normalize_state(state)
                            state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent_isolated.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        dx, dy, dz = agent_isolated.decode_action(action)
                        next_state = clip_state((state[0] + dx, state[1] + dy, state[2] + dz))

                        triggered = execute_Tr(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)

                        if target_path.issubset(triggered):
                            reward += 2.0

                        done = (step_count == N_STEPS - 1)

                        td_error = agent_isolated.store_transition(state, action, reward, next_state, done)
                        enhanced_standard_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                      next_state[0], next_state[1], next_state[2])

                        episode_similarities.append(jaccard_similarity(triggered, target_path))
                        episode_td_errors.append(td_error)

                        if prev_reward is not None:
                            enhanced_standard_metrics.record_action_improvement(reward, prev_reward)

                        prev_state = state
                        prev_triggered = triggered
                        prev_reward = reward
                        state = next_state
                        step_count += 1
                        episode_reward += reward
                        global_steps += 1

                        if global_steps % update_target_every == 0:
                            agent_isolated.update_target_model()

            if stage1_samples:
                for sample_idx in range(N_SAMPLES_STAGE1):
                    if sample_idx >= len(stage1_samples):
                        break

                    # stage1_samples
                    stage1_state_tuple, _, _, _ = stage1_samples[sample_idx]
                    step_count = 0
                    state = stage1_state_tuple  # 
                    prev_state = None
                    prev_triggered = None
                    prev_reward = None

                    for step in range(N_STEPS):
                        legal_actions = []
                        for a in range(agent_isolated.action_dim):
                            dx, dy, dz = agent_isolated.decode_action(a)
                            cand_next = (state[0] + dx, state[1] + dy, state[2] + dz)
                            if is_valid_state(cand_next):
                                legal_actions.append(a)

                        if not legal_actions:
                            break

                        if random.random() < agent_isolated.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            # 
                            norm_state = normalize_state(state)
                            state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent_isolated.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        dx, dy, dz = agent_isolated.decode_action(action)
                        next_state = clip_state((state[0] + dx, state[1] + dy, state[2] + dz))

                        triggered = execute_Tr(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                        reward *= 0.8

                        done = (step_count == N_STEPS - 1)

                        td_error = agent_isolated.store_transition(state, action, reward, next_state, done)
                        enhanced_standard_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                      next_state[0], next_state[1], next_state[2])

                        episode_similarities.append(jaccard_similarity(triggered, target_path))
                        episode_td_errors.append(td_error)

                        if prev_reward is not None:
                            enhanced_standard_metrics.record_action_improvement(reward, prev_reward)

                        prev_state = state
                        prev_triggered = triggered
                        prev_reward = reward
                        state = next_state
                        step_count += 1
                        episode_reward += reward
                        global_steps += 1

                        if global_steps % update_target_every == 0:
                            agent_isolated.update_target_model()

            for replay_round in range(REPLAY_TIMES):
                if len(agent_isolated.replay_buffer) >= batch_size:
                    agent_isolated.train(batch_size)

            trained_paths.add(path_idx)

        avg_similarity = np.mean(episode_similarities) if episode_similarities else 0
        avg_td_error = np.mean(episode_td_errors) if episode_td_errors else 0
        enhanced_standard_metrics.record_episode_metrics(episode, episode_reward, avg_similarity, avg_td_error,
                                                         agent_isolated.epsilon, "isolated")

        if episode % 10 == 0:
            agent_isolated.update_target_model()

        if len(trained_paths) == len(isolated_group):
            break

    return agent_isolated


# === Excel ===
def append_metrics_to_combined_excel(metrics_collector, agent_similar, agent_isolated, similar_group, isolated_group,
                                     targetPaths, filepath, run_number):
    """Metricfinal samplesExcelsheet"""

    # ===== Metric =====
    training_time = metrics_collector.end_time - metrics_collector.start_time if metrics_collector.end_time else 0
    avg_memory = metrics_collector.total_memory_usage / metrics_collector.memory_check_count if metrics_collector.memory_check_count > 0 else 0
    avg_similarity = np.mean(
        metrics_collector.final_output_similarities) if metrics_collector.final_output_similarities else 0
    avg_td_error = np.mean(metrics_collector.td_errors) if metrics_collector.td_errors else 0
    action_improve_rate = np.mean(metrics_collector.action_improvements) if metrics_collector.action_improvements else 0

    performance_row = {
        'Run': run_number,
        'Average Similarity': f"{avg_similarity:.4f}",
        'TD Error': f"{avg_td_error:.4f}",
        'Action Improve Rate': f"{action_improve_rate:.4f}",
        'Memory(MB)': f"{avg_memory:.2f}"
    }

    # ===== final samples =====
    sample_rows = []

    # 
    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_similar.replay_buffer.get_high_reward_samples(target_path, num_samples=20)

        for state_tuple, reward, sim, triggered in high_reward_samples:
            sample_rows.append({
                'Run': run_number,
                'Group Type': 'Similar',  #
                'Path ID': path_idx + 1,
                'X': state_tuple[0],
                'Y': state_tuple[1],
                'Z': state_tuple[2],
                'Similarity': f"{sim:.4f}",
                'Reward': f"{reward:.2f}",  #
                'Triggered Count': len(triggered),  #
                'Target Count': len(target_path),  #
                'Triggered Rules': str(sorted(triggered)),  #
                'Target Rules': str(sorted(target_path))  #
            })
    # 
    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)

        for state_tuple, reward, sim, triggered in high_reward_samples:
            sample_rows.append({
                'Run': run_number,
                'Group Type': 'Isolated',  # 修复此处
                'Path ID': path_idx + 1,
                'X': state_tuple[0],
                'Y': state_tuple[1],
                'Z': state_tuple[2],
                'Similarity': f"{sim:.4f}",
                'Reward': f"{reward:.2f}",  #
                'Triggered Count': len(triggered),  #
                'Target Count': len(target_path),  #
                'Triggered Rules': str(sorted(triggered)),  #
                'Target Rules': str(sorted(target_path))  #
            })

    # ===== Excel =====
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    # 
    if os.path.exists(filepath):
        # Metricsheet
        try:
            df_performance = pd.read_excel(filepath, sheet_name='Metric')
            df_performance = pd.concat([df_performance, pd.DataFrame([performance_row])], ignore_index=True)
        except:
            df_performance = pd.DataFrame([performance_row])

        # sheet
        try:
            df_samples = pd.read_excel(filepath, sheet_name='final samples')
            df_samples = pd.concat([df_samples, pd.DataFrame(sample_rows)], ignore_index=True)
        except:
            df_samples = pd.DataFrame(sample_rows)
    else:
        df_performance = pd.DataFrame([performance_row])
        df_samples = pd.DataFrame(sample_rows)

    # ===== Excel(sheet) =====
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Metricsheet
        df_performance.to_excel(writer, sheet_name='Metric', index=False)

        # sheet
        df_samples.to_excel(writer, sheet_name='final samples', index=False)

        workbook = writer.book

        # ===== Metricsheet =====
        ws_performance = writer.sheets['Metric']

        # 
        ws_performance.column_dimensions['A'].width = 15
        for col in ['B', 'C', 'D', 'E']:
            ws_performance.column_dimensions[col].width = 20

        # 
        header_font = Font(bold=True, size=11)
        for cell in ws_performance[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 
        for row in ws_performance.iter_rows(min_row=2, max_row=ws_performance.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # ===== sheet =====
        ws_samples = writer.sheets['final samples']

        # 
        column_widths = {
            'A': 12, 'B': 15, 'C': 12, 'D': 10, 'E': 10, 'F': 10,
            'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 40, 'L': 40
        }
        for col, width in column_widths.items():
            ws_samples.column_dimensions[col].width = width

        # 
        for cell in ws_samples[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 
        for row in ws_samples.iter_rows(min_row=2, max_row=ws_samples.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    print(f"Run {run_number} run: {filepath}")
    print(f"  - Metricsheet: ")
    print(f"  - final samplessheet: ")


# ===  run ===
def run_single_experiment(run_number, results_save_dir):
    """"""
    print(f"\n{'=' * 80}")
    print(f"Start run  {run_number}  run")
    print(f"{'=' * 80}\n")

    # Metric
    enhanced_standard_metrics.reset()
    enhanced_standard_metrics.start_training()

    model_path_similar = os.path.join(results_save_dir, f"run{run_number}_similar.pth")
    model_path_isolated = os.path.join(results_save_dir, f"run{run_number}_isolated.pth")
    path_documents = r"D:\Experiment\CNN\DQNNEW\path_samples_enhanced_standard"

    similar_group, isolated_group = group_paths_by_similarity(targetPaths)

    # Run : Path 
    if run_number == 1:
        generate_samples_for_similar_paths(similar_group, num_total=2000, top_k=200)

    replay_buffer = StandardExperienceReplay(capacity=10000)
    state_dim = 3
    action_dim = 30
    agent = StandardDQNAgent(state_dim, action_dim, replay_buffer)

    agent = enhanced_standard_generate_and_train_for_similar_paths(agent, similar_group, path_documents, episodes=500,
                                                                   batch_size=32, is_isolated=False)

    os.makedirs(os.path.dirname(model_path_similar), exist_ok=True)
    torch.save({
        'model_state_dict': agent.model.state_dict(),
        'optimizer_state_dict': agent.optimizer.state_dict(),
        'epsilon': agent.epsilon
    }, model_path_similar)

    # Run : Path 
    if run_number == 1:
        generate_samples_for_isolated_paths_standard(agent, isolated_group, num_total=2000, top_k=200)

    # Run : Path 
    isolated_replay_buffer = StandardExperienceReplay(capacity=15000)
    agent_isolated = StandardDQNAgent(state_dim, action_dim, isolated_replay_buffer)

    try:
        checkpoint = torch.load(model_path_similar)
        agent_isolated.model.load_state_dict(checkpoint['model_state_dict'])
        agent_isolated.target_model.load_state_dict(checkpoint['model_state_dict'])
        agent_isolated.epsilon = checkpoint.get('epsilon', 0.5)
    except Exception as e:
        pass

    agent_isolated = enhanced_standard_generate_and_train_for_isolated_paths(
        agent_similar=agent,
        agent_isolated=agent_isolated,
        similar_group=similar_group,
        isolated_group=isolated_group,
        path_documents=path_documents,
        episodes=500,
        batch_size=32,
        is_isolated=True
    )

    os.makedirs(os.path.dirname(model_path_isolated), exist_ok=True)
    torch.save({
        'model_state_dict': agent_isolated.model.state_dict(),
        'optimizer_state_dict': agent_isolated.optimizer.state_dict(),
        'epsilon': agent_isolated.epsilon
    }, model_path_isolated)

    # 
    enhanced_standard_metrics.end_training()

    # final samples
    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, _, sim, triggered in high_reward_samples:
            enhanced_standard_metrics.record_final_output_sample(triggered, target_path)

    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, _, sim, triggered in high_reward_samples:
            enhanced_standard_metrics.record_final_output_sample(triggered, target_path)

    # ===== : Excel =====
    combined_excel_path = os.path.join(results_save_dir, "__.xlsx")

    append_metrics_to_combined_excel(
        metrics_collector=enhanced_standard_metrics,
        agent_similar=agent,
        agent_isolated=agent_isolated,
        similar_group=similar_group,
        isolated_group=isolated_group,
        targetPaths=targetPaths,
        filepath=combined_excel_path,
        run_number=run_number
    )

    #  runMetric
    avg_similarity = np.mean(enhanced_standard_metrics.final_output_similarities)
    print(f"\nRun  {run_number}  runcompleted:")
    print(f"  Average Similarity: {avg_similarity:.4f}")
    print(f"  : {enhanced_standard_metrics.step_count}")


if __name__ == "__main__":
    results_save_dir = r"D:\Experiment\CNN\_"  # 
    os.makedirs(results_save_dir, exist_ok=True)

    # 20
    NUM_RUNS = 20

    print("=" * 80)
    print(f" {NUM_RUNS} DQN()")
    print(f": X[{X_MIN}, {X_MAX}], Y[{Y_MIN}, {Y_MAX}], Z[{Z_MIN}, {Z_MAX}]")
    print("=" * 80)

    for run in range(1, NUM_RUNS + 1):
        try:
            run_single_experiment(run, results_save_dir)
        except Exception as e:
            print(f"\nRun  {run}  run: {str(e)}")
            import traceback

            traceback.print_exc()
            continue

    print("\n" + "=" * 80)
    print(f" {NUM_RUNS}  runcompleted")
    print(f": {results_save_dir}")
    print("=" * 80)