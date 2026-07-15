import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === 全局取值范围配置 ===
STATE_MIN_W, STATE_MAX_W = 2, 100
STATE_MIN_T, STATE_MAX_T = 2, 100
STATE_MIN_Z, STATE_MAX_Z = 2, 100
# === / ===
class StateNormalizer:
    """: [0,1]"""
    def __init__(self):
        self.ranges = {
            'weather': (STATE_MIN_W, STATE_MAX_W),
            'time_period': (STATE_MIN_T, STATE_MAX_T),
            'z': (STATE_MIN_Z, STATE_MAX_Z)
        }

    def normalize(self, state):
        """[0,1]"""
        state = np.array(state, dtype=np.float32)
        normalized = np.zeros_like(state)

        # 
        normalized[0] = (state[0] - self.ranges['weather'][0]) / (
                    self.ranges['weather'][1] - self.ranges['weather'][0])  # weather
        normalized[1] = (state[1] - self.ranges['time_period'][0]) / (
                    self.ranges['time_period'][1] - self.ranges['time_period'][0])  # time_period
        normalized[2] = (state[2] - self.ranges['z'][0]) / (self.ranges['z'][1] - self.ranges['z'][0])  # z

        return normalized

    def denormalize(self, normalized_state):
        """"""
        normalized_state = np.array(normalized_state, dtype=np.float32)
        denormalized = np.zeros_like(normalized_state)

        # 
        denormalized[0] = normalized_state[0] * (self.ranges['weather'][1] - self.ranges['weather'][0]) + \
                          self.ranges['weather'][0]  # weather
        denormalized[1] = normalized_state[1] * (self.ranges['time_period'][1] - self.ranges['time_period'][0]) + \
                          self.ranges['time_period'][0]  # time_period
        denormalized[2] = normalized_state[2] * (self.ranges['z'][1] - self.ranges['z'][0]) + self.ranges['z'][0]  # z

        # 
        denormalized[0] = np.clip(np.round(denormalized[0]), self.ranges['weather'][0],
                                  self.ranges['weather'][1]).astype(int)
        denormalized[1] = np.clip(np.round(denormalized[1]), self.ranges['time_period'][0],
                                  self.ranges['time_period'][1]).astype(int)
        denormalized[2] = np.clip(np.round(denormalized[2]), self.ranges['z'][0], self.ranges['z'][1]).astype(int)

        return denormalized


# 
normalizer = StateNormalizer()


# === reward function ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward


def execute_Tr(a):
    """
    替换原有的编排规则函数
    参数a: 包含3个元素的元组或数组，分别对应path_depth, file_count, access_level
    返回: 触发的规则编号集合
    """
    path_depth, file_count, access_level = float(a[0]), int(a[1]), float(a[2])

    # 使用正确的变量名
    x, y, z = path_depth, file_count, access_level

    triggered = set()
    # 创建一个字典来存储b数组的值，用于跟踪哪些规则被触发
    b = {}
    x_score = max(0, 100 - x * 1.67)

    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[0] = 1
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((z * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[1] = 2
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * x) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[2] = 3
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[3] = 4
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * 43) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[4] = 5
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((74 * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[5] = 6
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (47 + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[6] = 7
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (x + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[7] = 8
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (y + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[8] = 9
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 10) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[9] = 10
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) != 50 and x ** 2 + y ** 2 > z ** 2):
        b[10] = 11
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 29 and x ** 2 + y ** 2 > z ** 2):
        b[11] = 12
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 69 and x ** 2 + y ** 2 > z ** 2):
        b[12] = 13
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2):
        b[13] = 14
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and y ** 2 + y ** 2 > z ** 2):
        b[14] = 15
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 3 + y ** 2 > z ** 2):
        b[15] = 16
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + x ** 2 > z ** 2):
        b[16] = 17
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 3 > z ** 2):
        b[17] = 18
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 != z ** 2):
        b[18] = 19
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2.5):
        b[19] = 20
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 1.4 > z ** 2):
        b[20] = 21
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + 34 ** 2 > z ** 2):
        b[21] = 22
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and 64 ** 2 + y ** 2 > z ** 2):
        b[22] = 23
    if (x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2:
        pattern_type = "311A1"
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((z ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[23] = 24
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - z ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[24] = 25
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 1.5) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[25] = 26
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2.3 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[26] = 27
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((45 ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[27] = 28
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - 31 ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[28] = 29
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[29] = 30
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (x + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[30] = 31
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 2) < -30 or (abs(x - y) * z) / 100 > 45):
        b[31] = 32
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (26 + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[32] = 33
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y - 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[33] = 34
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) != -30 or (abs(x - y) * z) / 100 > 45):
        b[34] = 35
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -10 or (abs(x - y) * z) / 100 > 45):
        b[35] = 36
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 and (abs(x - y) * z) / 100 > 45):
        b[36] = 37
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(z - y) * z) / 100 > 45):
        b[37] = 38
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - z) * z) / 100 > 45):
        b[38] = 39
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * y) / 100 > 45):
        b[39] = 40
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * x) / 100 > 45):
        b[40] = 41
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * 34) / 100 > 45):
        b[41] = 42
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - 26) * z) / 100 > 45):
        b[42] = 43
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(52 - y) * z) / 100 > 45):
        b[43] = 44
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 76 > 45):
        b[44] = 45
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 != 45):
        b[45] = 46
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 55):
        b[46] = 47
    if (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45:
        pattern_type = "311A2"

    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + x) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35):
        b[47] = 48
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((y + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35):
        b[48] = 49
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((24 + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35):
        b[49] = 50
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + 31) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35):
        b[50] = 51
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 1.5) ** 2 < z * 20 and (x * x * z) / 1000 < 35):
        b[51] = 52
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2.4 < z * 20 and (x * x * z) / 1000 < 35):
        b[52] = 53
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 != z * 20 and (x * x * z) / 1000 < 35):
        b[53] = 54
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < 25 * 20 and (x * x * z) / 1000 < 35):
        b[54] = 55
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 14 and (x * x * z) / 1000 < 35):
        b[55] = 56
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < x * 20 and (x * x * z) / 1000 < 35):
        b[56] = 57
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < y * 20 and (x * x * z) / 1000 < 35):
        b[57] = 58
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 1.2 < z * 20 and (x * x * z) / 1000 < 35):
        b[58] = 59
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 or (x * x * z) / 1000 < 35):
        b[59] = 60
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (y * x * z) / 1000 < 35):
        b[60] = 61
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (z * x * z) / 1000 < 35):
        b[61] = 62
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * 35 * z) / 1000 < 35):
        b[62] = 63
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * x * x) / 1000 < 35):
        b[63] = 64
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * x * y) / 1000 < 35):
        b[64] = 65
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * x * 26) / 1000 < 35):
        b[65] = 66
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * 54 * z) / 1000 < 35):
        b[66] = 67
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1050 < 35):
        b[67] = 68
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 != 35):
        b[68] = 69
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 25):
        b[69] = 70
    if ((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35:
        pattern_type = "311A3"
    if (x < 35 and y < 50) != (x < 35 or y < 50):
        b[70] = 71
    if (x < 35 and y < 50) != (x != 35 and y < 50):
        b[71] = 72
    if (x < 35 and y < 50) != (x < 19 and y < 50):
        b[72] = 73
    if (x < 35 and y < 50) != (x < 35 and y != 50):
        b[73] = 74
    if (x < 35 and y < 50) != (x < 35 and y < 61):
        b[74] = 75
    if x < 35 and y < 50:
        pattern_type = 312
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y < 35 or z < 65 and x < 50)):
        b[75] = 76
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y != 35 and z < 65 and x < 50)):
        b[76] = 77
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y < 25 and z < 65 and x < 50)):
        b[77] = 78
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y < 35 and z != 65 and x < 50)):
        b[78] = 79
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y < 35 and z < 72 and x < 50)):
        b[79] = 80
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y < 35 and z < 65 or x < 50)):
        b[80] = 81
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y < 35 and z < 65 and x != 50)):
        b[81] = 82
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 31)):
        b[82] = 83
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y < 35 and z < 53 and x < 50)):
        b[83] = 84
    if (y < 30 and z < 60) or (y < 35 and z < 65 and x < 50):
        pattern_type = 313
    if (z < 40 and x > 50) != (z < 40 or x > 50):
        b[84] = 85
    if (z < 40 and x > 50) != (z < 40 and x != 50):
        b[85] = 86
    if (z < 40 and x > 50) != (z != 40 and x > 50):
        b[86] = 87
    if (z < 40 and x > 50) != (z < 29 and x > 50):
        b[87] = 88
    if (z < 40 and x > 50) != (z < 40 and x > 62):
        b[88] = 89
    if z < 40 and x > 50:
        pattern_type = 314
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (y * y) / 100 < 28) or (x < 60 and y < 55)):
        b[89] = 90
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * x) / 100 < 28) or (x < 60 and y < 55)):
        b[90] = 91
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * 53) / 100 < 28) or (x < 60 and y < 55)):
        b[91] = 92
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (41 * y) / 100 < 28) or (x < 60 and y < 55)):
        b[92] = 93
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 88 < 28) or (x < 60 and y < 55)):
        b[93] = 94
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 100 != 28) or (x < 60 and y < 55)):
        b[94] = 95
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 100 < 21) or (x < 60 and y < 55)):
        b[95] = 96
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) and (x < 60 and y < 55)):
        b[96] = 97
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 or y < 55)):
        b[97] = 98
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x != 60 and y < 55)):
        b[98] = 99
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 69 and y < 55)):
        b[99] = 100
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y != 55)):
        b[100] = 101
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 45)):
        b[101] = 102
    if (40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55):
        pattern_type = 315
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((x ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)):
        b[102] = 103
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((z ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)):
        b[103] = 104
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2.4) / (x + z + 1) < 22) or (y < 65 and x < 70)):
        b[104] = 105
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (z + z + 1) < 22) or (y < 65 and x < 70)):
        b[105] = 106
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (y + z + 1) < 22) or (y < 65 and x < 70)):
        b[106] = 107
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + x + 1) < 22) or (y < 65 and x < 70)):
        b[107] = 108
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + y + 1) < 22) or (y < 65 and x < 70)):
        b[108] = 109
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 7) < 22) or (y < 65 and x < 70)):
        b[109] = 110
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 1) != 22) or (y < 65 and x < 70)):
        b[110] = 111
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 1) < 32) or (y < 65 and x < 70)):
        b[111] = 112
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 1) < 22) and (y < 65 and x < 70)):
        b[112] = 113
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 1) < 22) or (y != 65 and x < 70)):
        b[113] = 114
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 1) < 22) or (y < 45 and x < 70)):
        b[114] = 115
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 1) < 22) or (y < 65 or x < 70)):
        b[115] = 116
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x != 70)):
        b[116] = 117
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 55)):
        b[117] = 118
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((24 ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)):
        b[118] = 119
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (56 + z + 1) < 22) or (y < 65 and x < 70)):
        b[119] = 120
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + 73 + 1) < 22) or (y < 65 and x < 70)):
        b[120] = 121
    if ((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70):
        pattern_type = 316
    if (50 <= y < 75 and x > 65) != (50 <= y < 75 or x > 65):
        b[121] = 122
    if (50 <= y < 75 and x > 65) != (50 <= y < 75 and x != 65):
        b[122] = 123
    if (50 <= y < 75 and x > 65) != (50 <= y < 75 and x > 47):
        b[123] = 124
    if (50 <= y < 75 and x > 65) != (50 <= y < 82 and x > 65):
        b[124] = 125
    if (50 <= y < 75 and x > 65) != (57 <= y < 75 and x > 65):
        b[125] = 126
    if 50 <= y < 75 and x > 65:
        pattern_type = 317
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 89 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)):
        b[126] = 127
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((41 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)):
        b[127] = 128
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 or y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)):
        b[128] = 129
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y != 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)):
        b[129] = 130
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 79 and (z * y) / 100 < 52) or (z < 75 and y > 75)):
        b[130] = 131
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 or (z * y) / 100 < 52) or (z < 75 and y > 75)):
        b[131] = 132
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (y * y) / 100 < 52) or (z < 75 and y > 75)):
        b[132] = 133
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * z) / 100 < 52) or (z < 75 and y > 75)):
        b[133] = 134
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * 26) / 100 < 52) or (z < 75 and y > 75)):
        b[134] = 135
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (53 * y) / 100 < 52) or (z < 75 and y > 75)):
        b[135] = 136
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * y) / 89 < 52) or (z < 75 and y > 75)):
        b[136] = 137
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * y) / 100 != 52) or (z < 75 and y > 75)):
        b[137] = 138
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * y) / 100 < 39) or (z < 75 and y > 75)):
        b[138] = 139
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) and (z < 75 and y > 75)):
        b[139] = 140
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z != 75 and y > 75)):
        b[140] = 141
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 65 and y > 75)):
        b[141] = 142
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 or y > 75)):
        b[142] = 143
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y != 75)):
        b[143] = 144
    if (55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75):
        pattern_type = 318
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x != 88 and y > 75) or (78 <= x < 90 and y > 72)):
        b[144] = 145
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x < 94 and y > 75) or (78 <= x < 90 and y > 72)):
        b[145] = 146
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x < 88 or y > 75) or (78 <= x < 90 and y > 72)):
        b[146] = 147
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x < 88 and y != 75) or (78 <= x < 90 and y > 72)):
        b[147] = 148
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x < 88 and y > 67) or (78 <= x < 90 and y > 72)):
        b[148] = 149
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x < 88 and y > 75) and (78 <= x < 90 and y > 72)):
        b[149] = 150
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x < 88 and y > 75) or (78 <= x < 90 or y > 72)):
        b[150] = 151
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y != 72)):
        b[151] = 152
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 62)):
        b[152] = 153
    if (75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72):
        pattern_type = 319
    if (x >= 88 and y >= 85 and z >= 85) != (x >= 88 or y >= 85 and z >= 85):
        b[153] = 154
    if (x >= 88 and y >= 85 and z >= 85) != (x != 88 and y >= 85 and z >= 85):
        b[154] = 155
    if (x >= 88 and y >= 85 and z >= 85) != (x >= 78 and y >= 85 and z >= 85):
        b[155] = 156
    if (x >= 88 and y >= 85 and z >= 85) != (x >= 88 and y >= 85 or z >= 85):
        b[156] = 157
    if (x >= 88 and y >= 85 and z >= 85) != (x >= 88 and y != 85 and z >= 85):
        b[157] = 158
    if (x >= 88 and y >= 85 and z >= 85) != (x >= 88 and y >= 85 and z != 85):
        b[158] = 159
    if x >= 88 and y >= 85 and z >= 85:
        pattern_type = 320

    # 返回被触发的规则编号集合
    return set(b.values())

targetPaths = [
    {1, 6, 11, 14, 25, 26, 29, 37, 48, 51, 54, 58, 59, 60, 71, 74, 76, 77, 81, 85, 86, 98, 101, 103, 111, 114, 116, 119,
     122, 123, 129, 132, 143, 144, 151},
    {1, 4, 5, 6, 11, 14, 25, 26, 29, 37, 54, 59, 60, 71, 74, 76, 77, 81, 85, 86, 98, 101, 103, 104, 111, 114, 116, 119,
     122, 123, 129, 132, 143, 144, 151},
    {1, 2, 6, 8, 11, 14, 24, 26, 29, 37, 48, 51, 54, 59, 60, 71, 74, 76, 77, 81, 98, 101, 103, 111, 114, 116, 119, 122,
     123, 131, 133, 137, 139, 140, 151},
    {1, 2, 7, 8, 9, 10, 13, 24, 35, 46, 54, 60, 71, 72, 98, 99, 104, 108, 109, 110, 111, 112, 116, 117, 119, 121, 122,
     132, 143, 144, 147, 148, 151, 152},
    {1, 3, 4, 5, 6, 11, 14, 25, 26, 27, 29, 37, 54, 60, 71, 74, 76, 77, 81, 85, 86, 98, 101, 104, 111, 114, 116, 119,
     122, 123, 129, 132, 143, 144, 151},
    {1, 2, 6, 8, 11, 14, 24, 26, 29, 37, 48, 51, 54, 59, 60, 71, 74, 81, 98, 101, 103, 111, 114, 116, 119, 129, 132,
     135, 136, 138, 141, 143, 147, 151},
    {1, 2, 6, 11, 14, 49, 50, 52, 53, 55, 56, 57, 71, 74, 81, 98, 101, 103, 111, 114, 116, 119, 122, 123, 129, 132, 134,
     135, 136, 138, 143, 144, 151},
    {1, 2, 6, 11, 14, 24, 26, 29, 37, 48, 51, 54, 58, 59, 60, 71, 74, 76, 77, 81, 98, 101, 103, 111, 114, 116, 119, 122,
     123, 128, 129, 132, 143, 144},
    {1, 2, 6, 11, 14, 24, 26, 29, 37, 49, 50, 52, 53, 55, 56, 57, 71, 74, 76, 77, 81, 98, 101, 103, 111, 116, 119, 122,
     123, 129, 130, 132, 143, 144},
    {1, 3, 4, 5, 6, 11, 12, 14, 24, 28, 30, 31, 33, 34, 35, 36, 46, 54, 60, 89, 98, 101, 104, 111, 114, 116, 119, 122,
     123, 124, 129, 132, 143, 144},
    {1, 4, 5, 15, 20, 23, 37, 38, 39, 40, 42, 44, 47, 71, 72, 76, 85, 87, 98, 99, 103, 104, 113, 122, 129, 130, 132,
     143, 144, 147, 148, 151, 152},
    {3, 5, 9, 10, 13, 17, 20, 21, 22, 24, 26, 27, 29, 37, 85, 87, 98, 101, 103, 111, 114, 116, 119, 129, 132, 135, 136,
     138, 141, 143, 147, 151},
    {1, 4, 15, 20, 23, 37, 38, 39, 40, 42, 43, 44, 71, 72, 76, 85, 87, 98, 99, 103, 104, 113, 122, 129, 130, 132, 143,
     144, 147, 148, 151, 152},
    {1, 2, 6, 11, 14, 24, 26, 28, 29, 37, 49, 50, 52, 53, 55, 56, 57, 71, 74, 75, 76, 77, 81, 98, 113, 115, 122, 123,
     129, 130, 132, 143, 144},
    {24, 26, 28, 29, 37, 49, 50, 52, 53, 55, 56, 57, 58, 61, 62, 63, 67, 71, 74, 75, 81, 97, 102, 113, 115, 122, 123,
     129, 130, 132, 143, 144},
    {24, 26, 28, 29, 37, 49, 50, 52, 53, 55, 56, 57, 58, 61, 62, 63, 67, 70, 71, 74, 75, 81, 98, 113, 115, 122, 123,
     129, 130, 132, 143, 144},
    {1, 2, 6, 7, 8, 9, 10, 13, 24, 35, 46, 54, 60, 71, 72, 98, 99, 104, 108, 109, 110, 111, 112, 116, 117, 119, 121,
     122, 132, 143, 144, 154},
    {1, 2, 7, 8, 9, 10, 13, 24, 35, 46, 54, 60, 71, 72, 76, 81, 82, 98, 99, 103, 105, 106, 107, 113, 122, 132, 143, 144,
     147, 148, 151, 152},
    {1, 4, 15, 20, 35, 41, 45, 46, 71, 72, 76, 81, 82, 85, 87, 98, 99, 103, 104, 105, 113, 122, 129, 130, 132, 143, 144,
     147, 148, 151, 152},
    {1, 2, 7, 8, 9, 10, 13, 24, 35, 46, 54, 60, 98, 99, 104, 108, 109, 111, 112, 116, 117, 119, 121, 126, 132, 143, 144,
     147, 148, 151, 152},
    {14, 16, 17, 18, 19, 22, 37, 39, 40, 42, 43, 44, 47, 71, 72, 76, 85, 87, 98, 99, 103, 104, 113, 122, 129, 130, 132,
     147, 148, 151, 152},
    {1, 3, 4, 5, 6, 11, 12, 14, 24, 30, 33, 35, 36, 46, 54, 59, 60, 76, 77, 81, 85, 86, 90, 94, 96, 97, 113, 115, 122,
     123, 132, 143, 144},
    {1, 4, 5, 15, 20, 21, 22, 23, 25, 35, 38, 43, 45, 46, 71, 72, 85, 87, 98, 99, 103, 104, 105, 113, 122, 132, 151,
     152, 154, 157, 158},
    {3, 5, 13, 20, 21, 22, 24, 26, 27, 29, 30, 32, 37, 85, 87, 98, 101, 111, 114, 116, 119, 129, 132, 135, 136, 138,
     141, 143, 147, 151},
    {24, 26, 27, 28, 29, 37, 49, 50, 52, 53, 55, 56, 57, 58, 61, 62, 63, 67, 73, 76, 77, 81, 97, 102, 113, 115, 129,
     130, 132, 143, 144},
    {14, 16, 17, 19, 37, 39, 40, 42, 44, 60, 61, 65, 69, 71, 72, 76, 85, 87, 98, 99, 103, 104, 113, 122, 132, 147, 148,
     151, 152, 157},
    {2, 7, 8, 9, 10, 13, 24, 28, 30, 31, 33, 35, 36, 46, 88, 89, 98, 101, 104, 111, 114, 116, 119, 122, 123, 124, 129,
     132, 143, 144},
    {3, 5, 9, 13, 17, 20, 21, 22, 24, 26, 27, 29, 37, 85, 87, 98, 101, 103, 111, 114, 116, 119, 141, 143, 147, 151, 154,
     155, 157},
    {3, 4, 5, 6, 11, 12, 14, 24, 30, 35, 36, 46, 54, 59, 60, 89, 91, 92, 93, 95, 98, 101, 113, 115, 122, 123, 124, 132,
     143, 144},
    {1, 2, 6, 7, 8, 9, 10, 13, 24, 35, 46, 54, 60, 71, 72, 98, 99, 103, 105, 106, 107, 113, 120, 122, 132, 143, 144,
     147, 148},
    {1, 2, 6, 8, 11, 12, 14, 24, 26, 29, 37, 71, 74, 81, 98, 101, 103, 111, 114, 116, 119, 122, 123, 127, 131, 137, 139,
     140},
    {1, 3, 4, 5, 6, 12, 14, 24, 28, 30, 33, 35, 36, 46, 54, 59, 60, 89, 98, 101, 113, 115, 118, 122, 123, 124, 132, 143,
     144},
    {2, 6, 7, 8, 9, 10, 13, 24, 28, 30, 33, 35, 36, 46, 54, 60, 104, 111, 119, 122, 140, 145, 146, 147, 151, 154, 157,
     159},
    {3, 4, 5, 6, 11, 12, 14, 24, 35, 46, 54, 60, 89, 93, 95, 98, 99, 100, 113, 115, 118, 122, 123, 124, 132, 143, 144},
    {5, 20, 21, 22, 23, 28, 35, 43, 46, 85, 87, 111, 112, 119, 122, 125, 141, 143, 145, 146, 147, 151, 154, 157, 158},
    {5, 20, 28, 35, 46, 85, 87, 111, 112, 119, 129, 132, 133, 135, 136, 138, 143, 144, 147, 148, 149, 151, 152, 153},
    {5, 20, 24, 28, 33, 35, 36, 43, 46, 85, 87, 111, 119, 122, 129, 132, 135, 138, 141, 143, 150, 154, 157, 159},
    {1, 2, 4, 6, 7, 8, 11, 12, 14, 25, 26, 27, 29, 37, 81, 98, 101, 103, 111, 114, 116, 119, 140, 142, 147, 151},
    {24, 26, 27, 28, 29, 30, 37, 60, 64, 65, 66, 69, 71, 75, 81, 97, 102, 113, 115, 122, 123, 129, 130, 132},
    {6, 11, 14, 24, 25, 26, 27, 28, 29, 32, 37, 48, 54, 55, 58, 59, 60, 78, 85, 86, 97, 113, 132, 143, 144},
    {25, 31, 35, 36, 46, 53, 55, 57, 58, 61, 62, 63, 67, 70, 73, 76, 79, 81, 97, 129, 130, 132, 143, 144},
    {25, 35, 36, 46, 52, 53, 55, 56, 57, 58, 61, 62, 63, 67, 70, 73, 78, 84, 97, 129, 130, 132, 143, 144},
    {25, 35, 36, 46, 53, 55, 57, 58, 61, 62, 63, 67, 70, 73, 76, 79, 80, 81, 97, 129, 130, 132, 143, 144},
    {5, 20, 21, 22, 28, 30, 31, 33, 35, 36, 43, 46, 85, 87, 111, 119, 122, 141, 143, 154, 155, 156, 157},
    {24, 26, 28, 29, 30, 37, 60, 64, 65, 66, 68, 69, 71, 74, 75, 81, 98, 101, 113, 115, 122, 123, 157},
    {3, 4, 5, 6, 11, 12, 14, 24, 35, 46, 54, 59, 60, 71, 72, 78, 83, 85, 86, 97, 113, 132, 143, 144},

]

# 
target_paths = [set(path) for path in targetPaths]


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0


def compute_path_similarity_matrix(paths):
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix


def group_paths_by_similarity(paths):
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.mean(avg_sim_scores)

    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]
    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)

    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group


def compute_robustness(state, path):
    """()"""
    weather, time_period, z = state
    base = execute_Tr(weather, time_period, z)
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dw in [-1, 0, 1]:
        for dt in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dw == dt == dz == 0:
                    continue
                neighbor = np.clip(np.array(state) + np.array([dw, dt, dz]),
                                    [STATE_MIN_W, STATE_MIN_T, STATE_MIN_Z],
                                    [STATE_MAX_W, STATE_MAX_T, STATE_MAX_Z])
                neighbor = tuple(neighbor)
                nw, nt, nz = neighbor
                n_trig = execute_Tr(nw, nt, nz)
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def compute_q_value_score(state, similar_model):
    """Q()"""
    if similar_model is None:
        return 0.0

    try:
        # 
        normalized_state = normalizer.normalize(state)
        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = similar_model(state_tensor)
            max_q_value = torch.max(q_values).item()
            normalized_q = max_q_value / 20.0
            normalized_q = max(0.0, min(1.0, normalized_q))
            return 1.0 - normalized_q
    except:
        return 0.0


def generate_samples_for_similar_paths(similar_group, num_candidates=2000, top_k=200, run_id=1):
    SIMILAR_WEIGHTS = [0.55, 0.39, 0.06]

    def save_samples(path_id, samples, base_dir, group_type="similar"):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_{group_type}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"{group_type.title()} Group Path {path_id} - Run {run_id}\n")
            f.write("weather time_period z\tScore\tSimilarity\tLengthDiff\tRobustness\n")
            for s in samples:
                weather, time_period, z = s['state']
                f.write(
                    f"{weather} {time_period} {z}\t{s['score']:.4f}\t{s['similarity']:.4f}\t{s['length_diff']:.4f}\t{s['robustness']:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_grouped"

    for path_idx in similar_group:
        path = target_paths[path_idx]
        path_id = path_idx + 1
        candidate_samples = []
        attempts = 0

        while len(candidate_samples) < num_candidates and attempts < num_candidates * 10:
            attempts += 1

            weather = random.randint(STATE_MIN_W, STATE_MAX_W)
            time_period = random.randint(STATE_MIN_T, STATE_MAX_T)
            z = random.randint(STATE_MIN_Z, STATE_MAX_Z)
            state = (weather, time_period, z)

            triggered = execute_Tr(weather, time_period, z)

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)

            candidate_samples.append({
                'state': state,
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            })

        if candidate_samples:
            for sample in candidate_samples:
                score = (SIMILAR_WEIGHTS[0] * sample['similarity'] +
                         SIMILAR_WEIGHTS[1] * sample['length_diff'] +
                         SIMILAR_WEIGHTS[2] * sample['robustness'])
                sample['score'] = score

            candidate_samples.sort(key=lambda x: x['score'], reverse=True)
            selected_samples = candidate_samples[:top_k]
            save_samples(path_id=path_id, samples=selected_samples, base_dir=base_dir, group_type="similar")


def generate_samples_for_isolated_paths(isolated_group, similar_model, num_candidates=2000, top_k=200, run_id=1):
    ISOLATED_WEIGHTS = [0.18, 0.21, 0.32, 0.29]

    def save_samples(path_id, samples, base_dir, group_type="isolated"):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_{group_type}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"{group_type.title()} Group Path {path_id} - Run {run_id}\n")
            f.write("weather time_period z\tScore\tSimilarity\tLengthDiff\tRobustness\tQValueScore\n")
            for s in samples:
                weather, time_period, z = s['state']
                f.write(
                    f"{weather} {time_period} {z}\t{s['score']:.4f}\t{s['similarity']:.4f}\t{s['length_diff']:.4f}\t{s['robustness']:.4f}\t{s['q_value_score']:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_grouped"

    for path_idx in isolated_group:
        path = target_paths[path_idx]
        path_id = path_idx + 1
        candidate_samples = []
        attempts = 0

        while len(candidate_samples) < num_candidates and attempts < num_candidates * 10:
            attempts += 1

            weather = random.randint(STATE_MIN_W, STATE_MAX_W)
            time_period = random.randint(STATE_MIN_T, STATE_MAX_T)
            z = random.randint(STATE_MIN_Z, STATE_MAX_Z)
            state = (weather, time_period, z)

            triggered = execute_Tr(weather, time_period, z)

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            q_score = compute_q_value_score(state, similar_model)

            candidate_samples.append({
                'state': state,
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'q_value_score': q_score,
                'triggered': triggered
            })

        if candidate_samples:
            for sample in candidate_samples:
                score = (ISOLATED_WEIGHTS[0] * sample['similarity'] +
                         ISOLATED_WEIGHTS[1] * sample['length_diff'] +
                         ISOLATED_WEIGHTS[2] * sample['robustness'] +
                         ISOLATED_WEIGHTS[3] * sample['q_value_score'])
                sample['score'] = score

            candidate_samples.sort(key=lambda x: x['score'], reverse=True)
            selected_samples = candidate_samples[:top_k]
            save_samples(path_id=path_id, samples=selected_samples, base_dir=base_dir, group_type="isolated")


class GroupExperienceReplay:
    def __init__(self, capacity=20000):
        self.capacity = capacity
        self.buffer = deque(maxlen=self.capacity)
        self.priorities = deque(maxlen=self.capacity)
        self.sampled_indices = set()  # 

    def append(self, experience):
        self.buffer.append(experience)
        self.priorities.append(experience[-1])

    def sample(self, batch_size, alpha=0.6):
        priorities = np.array(self.priorities) ** alpha
        probabilities = priorities / np.sum(priorities)
        batch_indices = np.random.choice(len(self.buffer), batch_size, p=probabilities)
        batch = [self.buffer[idx] for idx in batch_indices]
        return batch, batch_indices, probabilities[batch_indices]

    def update_priorities(self, batch_indices, td_errors):
        for idx, td_error in zip(batch_indices, td_errors):
            if idx < len(self.priorities):
                self.priorities[idx] = max(td_error, 1e-6)

    def __len__(self):
        return len(self.buffer)

    def get_high_reward_samples(self, target_path, num_samples=20):
        """(, )"""
        if len(self.buffer) == 0:
            return []

        samples_with_recalculated_scores = []
        for idx, experience in enumerate(self.buffer):
            # 
            if idx in self.sampled_indices:
                continue

            # 
            normalized_state_tensor = experience[0]
            normalized_state = normalized_state_tensor.cpu().numpy().flatten()
            state_tuple = tuple(normalizer.denormalize(normalized_state))

            weather, time_period, z = state_tuple
            triggered = execute_Tr(weather, time_period, z)
            new_reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_recalculated_scores.append((idx, state_tuple, new_reward, sim, triggered))

        # 
        samples_with_recalculated_scores.sort(key=lambda x: x[2], reverse=True)

        # num_samples
        selected = samples_with_recalculated_scores[:num_samples]

        # 
        for item in selected:
            self.sampled_indices.add(item[0])

        # : (state_tuple, reward, sim, triggered)
        return [(s[1], s[2], s[3], s[4]) for s in selected]

    def reset_sampled_indices(self):
        """"""
        self.sampled_indices.clear()


def load_path_data(file_path):
    """Path ()"""
    path_data = []
    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines[2:]:
            parts = line.strip().split("\t")
            state = tuple(map(int, parts[0].split()))
            path_data.append(state)
    return path_data


class DQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, 128)
        self.fc2 = nn.Linear(128, 64)
        self.fc3 = nn.Linear(64, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)


class DQNAgentWithPER:
    def __init__(self, state_dim, action_dim, replay_buffer, gamma=0.99, epsilon=1.0, epsilon_decay=0.995,
                 epsilon_min=0.1, learning_rate=0.001, alpha=0.6, beta=0.4):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer
        self.alpha = alpha
        self.beta = beta

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        """()"""
        delta_values = [1, -1]  # , 
        dim = action_idx // 2
        delta_idx = action_idx % 2
        delta = delta_values[delta_idx]
        if dim == 0:
            return (delta, 0, 0)
        elif dim == 1:
            return (0, delta, 0)
        elif dim == 2:
            return (0, 0, delta)

    def act(self, normalized_state):
        """()"""
        if random.random() < self.epsilon:
            return random.randrange(self.action_dim)
        state = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state)
        return torch.argmax(q_values, dim=1).item()

    def store_transition(self, normalized_state, action, reward, normalized_next_state, done):
        """()"""
        state = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        next_state = torch.tensor(normalized_next_state, dtype=torch.float32).unsqueeze(0).to(device)

        with torch.no_grad():
            q_values = self.model(state)
            next_q_values = self.target_model(next_state)
            max_next_q_values = next_q_values.max(1)[0]
            target_q_values = reward + (self.gamma * max_next_q_values * (1 - done))
            td_error = torch.abs(q_values[0][action] - target_q_values).item()

        self.replay_buffer.append((state, action, reward, next_state, done, td_error))
        return td_error

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return

        batch, batch_indices, probabilities = self.replay_buffer.sample(batch_size, alpha=self.alpha)
        states, actions, rewards, next_states, dones, _ = zip(*batch)

        weights = (len(self.replay_buffer) * probabilities) ** (-self.beta)
        weights = weights / weights.max()
        weights = torch.tensor(weights, dtype=torch.float32).to(device)

        states = torch.tensor(np.array([s.cpu().numpy().flatten() for s in states]), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array([ns.cpu().numpy().flatten() for ns in next_states]),
                                   dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))

        td_errors = current_q_values - target_q_values
        weighted_loss = (td_errors.pow(2) * weights).mean()

        self.optimizer.zero_grad()
        weighted_loss.backward()
        self.optimizer.step()

        new_priorities = torch.abs(td_errors).detach().cpu().numpy()
        self.replay_buffer.update_priorities(batch_indices, new_priorities)

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())


def train_group(group_paths, path_documents, replay_buffer, batch_size=32, group_name="", pretrained_model=None):
    """()- 3 minutes"""
    state_dim = 3
    action_dim = 6  # 2*3 (2delta * 3)

    agent = DQNAgentWithPER(state_dim, action_dim, replay_buffer)

    if pretrained_model is not None:
        print(f"  {group_name}: ()...")
        agent.model.load_state_dict(pretrained_model.state_dict())
        agent.target_model.load_state_dict(pretrained_model.state_dict())
        print(f"  {group_name}: completed")

    path_rewards = {}

    print(f"Start training{group_name}, Included Paths: {[idx + 1 for idx in group_paths]}")
    start_time = time.time()

    # === 3 minutes ===
    BATCH_SIZE = 50  # 
    N_SAMPLES = 200  # 
    N_STEPS = 3  # 
    N_ROUNDS = 5  # 
    N_BATCHES = 4  # 

    replay_count = 0

    for path_idx in group_paths:
        file_path = os.path.join(path_documents,
                                 f"path{path_idx + 1}_{'similar' if group_name == '' else 'isolated'}.txt")
        if not os.path.exists(file_path):
            print(f"    : Path {path_idx + 1}, ")
            continue

        path_data = load_path_data(file_path)  # 
        target_path = target_paths[path_idx]

        if path_idx not in path_rewards:
            path_rewards[path_idx] = 0

        print(f"\n  Start training path  {path_idx + 1},  {N_ROUNDS} ")

        for round_idx in range(N_ROUNDS):
            print(f"    Path  {path_idx + 1} - Run  {round_idx + 1}/{N_ROUNDS} ")

            for batch_idx in range(N_BATCHES):
                batch_start = batch_idx * BATCH_SIZE
                batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)

                # , 
                if batch_start >= len(path_data):
                    print(f"       {batch_idx + 1}: , ")
                    break

                print(f"       {batch_idx + 1}/{N_BATCHES} ( {batch_start}-{batch_end})")

                for sample_idx in range(batch_start, batch_end):
                    if sample_idx >= len(path_data):
                        break

                    state = path_data[sample_idx]  # 
                    prev_state = None
                    prev_triggered = None

                    for step in range(N_STEPS):
                        # 
                        normalized_state = normalizer.normalize(state)

                        # 
                        legal_actions = []
                        for a in range(agent.action_dim):
                            dw, dt, dz = agent.decode_action(a)
                            # 
                            cand_next = tuple(np.clip(np.array(state) + np.array([dw, dt, dz]),
                                                      [STATE_MIN_W, STATE_MIN_T, STATE_MIN_Z],
                                                      [STATE_MAX_W, STATE_MAX_T, STATE_MAX_Z]))
                            legal_actions.append(a)

                        if not legal_actions:
                            break

                        # 
                        if random.random() < agent.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        # ()
                        dw, dt, dz = agent.decode_action(action)
                        next_state = tuple(np.clip(np.array(state) + np.array([dw, dt, dz]),
                                                   [STATE_MIN_W, STATE_MIN_T, STATE_MIN_Z],
                                                   [STATE_MAX_W, STATE_MAX_T, STATE_MAX_Z]))

                        # 
                        normalized_next_state = normalizer.normalize(next_state)

                        # ()
                        weather, time_period, z = next_state
                        triggered = execute_Tr(weather, time_period, z)
                        reward = compute_reward(next_state, target_path, triggered,
                                                prev_triggered, prev_state)
                        done = (step == N_STEPS - 1)

                        # ()
                        agent.store_transition(normalized_state, action, reward, normalized_next_state, done)

                        # 
                        prev_state = state
                        prev_triggered = triggered
                        state = next_state
                        path_rewards[path_idx] += reward

                # 
                if len(agent.replay_buffer) >= batch_size:
                    agent.train(batch_size)
                    replay_count += 1

                    if replay_count % 2 == 0:
                        agent.update_target_model()

            print(f"      Path  {path_idx + 1} - Run  {round_idx + 1} completed")

        print(f"  Path  {path_idx + 1}  {N_ROUNDS} All completed ")

    training_time = time.time() - start_time
    print(f"\n{group_name}completed!")
    print(f"  : Path completed{N_ROUNDS}()")
    print(f"  : {replay_count}")
    print(f"  : {training_time:.2f} seconds")
    print(f"  : {len(replay_buffer)}")

    return agent, path_rewards, training_time


def generate_and_train_grouped_paths_staged(path_documents, similar_group, isolated_group, batch_size=32, run_id=1):
    """()- 3 minutes"""
    print(f"\n===  {run_id}/20 (3 minutes) ===")
    similar_group_paths = [idx + 1 for idx in similar_group]
    isolated_group_paths = [idx + 1 for idx in isolated_group]

    print(f"Path : {similar_group_paths}")
    print(f"Path : {isolated_group_paths}")

    total_start_time = time.time()

    print(f"\n[1] ...")
    # Sample generation
    generate_samples_for_similar_paths(similar_group, num_candidates=2000, top_k=200, run_id=run_id)

    print(f"\n[2] (, {5})...")
    similar_replay_buffer = GroupExperienceReplay(capacity=20000)
    similar_agent, similar_path_rewards, similar_training_time = train_group(
        similar_group, path_documents, similar_replay_buffer, batch_size,
        group_name="", pretrained_model=None
    )

    print(f"\n[3] ...")
    generate_samples_for_isolated_paths(isolated_group, similar_agent.model,
                                        num_candidates=2000, top_k=200, run_id=run_id)

    print(f"\n[4] (, {5})...")
    isolated_replay_buffer = GroupExperienceReplay(capacity=20000)
    isolated_agent, isolated_path_rewards, isolated_training_time = train_group(
        isolated_group, path_documents, isolated_replay_buffer, batch_size,
        group_name="", pretrained_model=similar_agent.model
    )

    total_path_rewards = {**similar_path_rewards, **isolated_path_rewards}
    total_cumulative_reward = sum(total_path_rewards.values())
    total_training_time = time.time() - total_start_time

    print(f"\n===  {run_id}/20 completed, : {total_training_time:.2f} seconds ===")
    print(f": {similar_training_time:.2f} seconds")
    print(f": {isolated_training_time:.2f} seconds")
    print(f" - : {len(similar_replay_buffer)}, : {len(isolated_replay_buffer)}")

    return similar_agent, isolated_agent, similar_replay_buffer, isolated_replay_buffer, \
        total_cumulative_reward, total_path_rewards, total_training_time


def create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir):
    """Excel()"""
    os.makedirs(output_dir, exist_ok=True)

    similar_group_paths = [idx + 1 for idx in similar_group]
    isolated_group_paths = [idx + 1 for idx in isolated_group]

    wb = Workbook()

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    header_color = "4472C4"
    similar_group_color = "E2EFDA"
    isolated_group_color = "FCE4D6"
    stats_color = "FFF2CC"

    ws_paths = wb.active
    ws_paths.title = "Path "

    path_headers = ['Path ID', ''] + [f'Run {i}' for i in range(1, 21)] + ['Average Similarity', 'Maximum Similarity',
                                                                                    'Minimum Similarity', 'Standard deviation']
    for col, header in enumerate(path_headers, 1):
        cell = ws_paths.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_paths.row_dimensions[1].height = 30

    for path_id in range(1, len(target_paths) + 1):
        row = path_id + 1

        if path_id in similar_group_paths:
            group_type = "High-correlation path group"
            row_color = similar_group_color
        elif path_id in isolated_group_paths:
            group_type = "Low-correlation path group"
            row_color = isolated_group_color
        else:
            group_type = "Ungrouped"
            row_color = "FFFFFF"

        cell = ws_paths.cell(row=row, column=1, value=f"Path {path_id}")
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        cell.border = thin_border

        cell = ws_paths.cell(row=row, column=2, value=group_type)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        cell.border = thin_border

        path_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            sim = run_data['path_similarities'].get(path_id, {}).get('avg_similarity', 0.0)
            path_similarities.append(sim)

            cell = ws_paths.cell(row=row, column=3 + run_idx, value=round(sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        stats_values = [
            round(np.mean(path_similarities), 4),
            round(np.max(path_similarities), 4),
            round(np.min(path_similarities), 4),
            round(np.std(path_similarities), 4)
        ]

        for i, value in enumerate(stats_values):
            cell = ws_paths.cell(row=row, column=23 + i, value=value)
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
            cell.font = Font(bold=True, size=10)
            cell.border = thin_border

    ws_paths.column_dimensions['A'].width = 13
    ws_paths.column_dimensions['B'].width = 16
    for col in range(3, 23):
        ws_paths.column_dimensions[get_column_letter(col)].width = 10
    for col in range(23, 27):
        ws_paths.column_dimensions[get_column_letter(col)].width = 13

    # === 2:  ===
    ws_groups = wb.create_sheet("")

    # ("screening")
    group_headers = ['Group Name', 'Included Paths'] + [f'Run {i}' for i in range(1, 21)] + ['Average Similarity', 'Standard deviation']
    for col, header in enumerate(group_headers, 1):
        cell = ws_groups.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_groups.row_dimensions[1].height = 30

    row = 2

    # Similar path group
    cell = ws_groups.cell(row=row, column=1, value="High-correlation path group")
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    cell.border = thin_border

    cell = ws_groups.cell(row=row, column=2, value=','.join(map(str, similar_group_paths)))
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    cell.border = thin_border

    group_similarities = []
    for run_idx, run_data in enumerate(all_runs_data):
        group_sim = np.mean(
            [run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0) for p in similar_group_paths])
        group_similarities.append(group_sim)

        cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(group_sim, 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 
    cell = ws_groups.cell(row=row, column=23, value=round(np.mean(group_similarities), 4))
    cell.number_format = '0.0000'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border

    cell = ws_groups.cell(row=row, column=24, value=round(np.std(group_similarities), 4))
    cell.number_format = '0.0000'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border

    row += 1

    # Isolated path group
    if isolated_group_paths:
        cell = ws_groups.cell(row=row, column=1, value="Low-correlation path group")
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        cell.border = thin_border

        cell = ws_groups.cell(row=row, column=2, value=','.join(map(str, isolated_group_paths)))
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        cell.border = thin_border

        isolated_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            iso_sim = np.mean(
                [run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0) for p in isolated_group_paths])
            isolated_similarities.append(iso_sim)

            cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(iso_sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # 
        cell = ws_groups.cell(row=row, column=23, value=round(np.mean(isolated_similarities), 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border

        cell = ws_groups.cell(row=row, column=24, value=round(np.std(isolated_similarities), 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border

    # 
    ws_groups.column_dimensions['A'].width = 16
    ws_groups.column_dimensions['B'].width = 22
    for col in range(3, 23):
        ws_groups.column_dimensions[get_column_letter(col)].width = 10
    ws_groups.column_dimensions[get_column_letter(23)].width = 14
    ws_groups.column_dimensions[get_column_letter(24)].width = 12

    # === 3: Detailed Sample Data ===
    ws_samples = wb.create_sheet("Detailed Sample Data")

    # ("")
    sample_headers = ['Run', 'Path ID', 'Sample ID', 'Weather', 'Time_Period', 'Z', 'Similarity', 'Triggered Rule Set']
    for col, header in enumerate(sample_headers, 1):
        cell = ws_samples.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_samples.row_dimensions[1].height = 30

    sample_row = 2
    #  runPath 
    for run_idx, run_data in enumerate(all_runs_data, 1):
        for path_id in range(1, 15):
            samples = run_data['path_samples'].get(path_id, [])

            # Path 
            if path_id in similar_group_paths:
                path_color = similar_group_color
            elif path_id in isolated_group_paths:
                path_color = isolated_group_color
            else:
                path_color = "FFFFFF"

            for sample_idx, (state_tuple, reward, sim, triggered) in enumerate(samples, 1):
                weather, time_period, z = state_tuple
                triggered_str = ','.join(map(str, sorted(triggered)))

                # Run
                cell = ws_samples.cell(row=sample_row, column=1, value=f"Run {run_idx}")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=path_color, end_color=path_color, fill_type="solid")
                cell.border = thin_border

                # Path ID
                cell = ws_samples.cell(row=sample_row, column=2, value=f"Path {path_id}")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=path_color, end_color=path_color, fill_type="solid")
                cell.border = thin_border

                # Sample ID
                cell = ws_samples.cell(row=sample_row, column=3, value=sample_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                # Weather, Time_Period, Z
                for col_offset, value in enumerate([weather, time_period, z]):
                    cell = ws_samples.cell(row=sample_row, column=4 + col_offset, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

                # Similarity
                cell = ws_samples.cell(row=sample_row, column=7, value=round(sim, 4))
                cell.number_format = '0.0000'
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                # Triggered Rule Set
                cell = ws_samples.cell(row=sample_row, column=8, value=f"{{{triggered_str}}}")
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border

                sample_row += 1

    # 
    sample_widths = [13, 13, 11, 10, 12, 8, 12, 45]
    for i, width in enumerate(sample_widths, 1):
        ws_samples.column_dimensions[get_column_letter(i)].width = width

    # 
    output_path = os.path.join(output_dir, "20 run_3 minutes.xlsx")
    wb.save(output_path)
    print(f"\n Consolidated Excel report generated: {output_path}")


def run_20_times_training():
    """20(3 minutes)- """
    model_path_base = r"D:\Experiment\CNN\DQNNEW\saved_models_3min_version"
    path_documents = r"D:\Experiment\CNN\DQNNEW\path_samples_grouped"
    output_dir = r"D:\Experiment\CNN\ComparisonExperiment2\excel_reports_3min_version"

    os.makedirs(model_path_base, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    similar_group, isolated_group = group_paths_by_similarity(target_paths)

    similar_group_display = [idx + 1 for idx in similar_group]
    isolated_group_display = [idx + 1 for idx in isolated_group]

    print("=" * 60)
    print("20 - 3 minutes")
    print("=" * 60)
    print("Training-scale configuration:")
    print("   Per path: 5")
    print("   Per round: 4")
    print("   Per batch: 50")
    print("   Per sample: 3")
    print("   Sample generation: 2000candidates -> 200final samples")
    print("   : save model parameters only(optimized version)")
    print("=" * 60)
    print(f"\nAutomatic grouping results:")
    print(f"Similar path group: {similar_group_display}")
    print(f"Isolated path group: {isolated_group_display}")
    print("\n" + "=" * 60)

    all_runs_data = []
    total_start_time = time.time()

    for run_id in range(1, 21):
        print(f"\n{'=' * 60}")
        print(f"Start run  {run_id}/20  run")
        print(f"{'=' * 60}")

        similar_agent, isolated_agent, similar_buffer, isolated_buffer, total_cumulative_reward, path_rewards, training_time = \
            generate_and_train_grouped_paths_staged(path_documents, similar_group, isolated_group, batch_size=32,
                                                    run_id=run_id)

        # === : save model parameters only,  ===
        similar_model_path = os.path.join(model_path_base, f"similar_group_model_run_{run_id}.pth")
        isolated_model_path = os.path.join(model_path_base, f"isolated_group_model_run_{run_id}.pth")

        # , 
        torch.save(similar_agent.model.state_dict(), similar_model_path)
        torch.save(isolated_agent.model.state_dict(), isolated_model_path)

        print(f"[Run {run_id}] Model saved(optimized version - )")

        # 
        similar_buffer.reset_sampled_indices()
        isolated_buffer.reset_sampled_indices()

        run_data = {
            'run_id': run_id,
            'training_time': training_time,
            'total_reward': total_cumulative_reward,
            'path_rewards': path_rewards,
            'path_similarities': {},
            'path_samples': {}
        }

        all_similarities = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            path_id = path_idx + 1

            if path_id in similar_group_display:
                buffer = similar_buffer
            elif path_id in isolated_group_display:
                buffer = isolated_buffer
            else:
                continue

            high_reward_samples = buffer.get_high_reward_samples(target_path, num_samples=20)

            if high_reward_samples:
                similarities = [sim for _, _, sim, _ in high_reward_samples]
                run_data['path_similarities'][path_idx + 1] = {
                    'avg_similarity': np.mean(similarities),
                    'max_similarity': np.max(similarities),
                    'min_similarity': np.min(similarities),
                    'sample_count': len(similarities)
                }
                run_data['path_samples'][path_idx + 1] = high_reward_samples
                all_similarities.extend(similarities)
            else:
                run_data['path_similarities'][path_idx + 1] = {
                    'avg_similarity': 0.0,
                    'max_similarity': 0.0,
                    'min_similarity': 0.0,
                    'sample_count': 0
                }
                run_data['path_samples'][path_idx + 1] = []

        if all_similarities:
            run_data['overall_avg_similarity'] = np.mean(all_similarities)
            run_data['max_similarity'] = np.max(all_similarities)
            run_data['min_similarity'] = np.min(all_similarities)
        else:
            run_data['overall_avg_similarity'] = 0.0
            run_data['max_similarity'] = 0.0
            run_data['min_similarity'] = 0.0

        all_runs_data.append(run_data)

        print(f"[Run {run_id}] completed! Overall Average Similarity: {run_data['overall_avg_similarity']:.4f}")
        print(f"{'=' * 60}\n")

    total_time = time.time() - total_start_time

    print("\nGenerating consolidated Excel report...")
    create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir)

    print("\n" + "=" * 60)
    print("20All completed! - 3 minutes")
    print("=" * 60)
    print(f":")
    print(f"  Per path: 5 x 4 x 50 x 3 = 3000/Path ")
    print(f"  Sample generation: 2000candidates -> 200final samples")
    print(f"  : save model parameters only(optimized version)")
    print(f"  Total elapsed time: {total_time:.2f} seconds ({total_time / 60:.2f} minutes)")
    print(f"  Average elapsed time per run: {total_time / 20:.2f} seconds")
    print(f"\nAverage similarity statistics:")
    avg_similarities = [r['overall_avg_similarity'] for r in all_runs_data]
    print(f"  Overall average: {np.mean(avg_similarities):.4f}")
    print(f"  Maximum: {np.max(avg_similarities):.4f}")
    print(f"  Minimum: {np.min(avg_similarities):.4f}")
    print(f"  Standard deviation: {np.std(avg_similarities):.4f}")
    print(f"\nAll results have been saved to: {output_dir}")
    print("=" * 60)


if __name__ == "__main__":
    run_20_times_training()