
import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
import numpy as np
import random
import time
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# ===  ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'MIN_X': 1, 'MAX_X': 100,
    'MIN_Y': 1, 'MAX_Y': 100,
    'MIN_Z': 1, 'MAX_Z': 100,
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 5,
    'NUM_ROUNDS': 5,
    'REPLAY_BATCH_SIZE': 64,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'LEARNING_RATE': 3e-4,
    'NUM_RUNS': 20,  # 20 run
    'TOP_K_SAMPLES': 20,
    'REPLAY_BUFFER_CAPACITY': 20000,  # Path 
    'TARGET_PATHS': [
        {1, 6, 11, 14, 25, 26, 29, 37, 48, 51, 54, 58, 59, 60, 71, 74, 76, 77, 81, 85, 86, 98, 101, 103, 111, 114, 116,
         119,
         122, 123, 129, 132, 143, 144, 151},
        {1, 4, 5, 6, 11, 14, 25, 26, 29, 37, 54, 59, 60, 71, 74, 76, 77, 81, 85, 86, 98, 101, 103, 104, 111, 114, 116,
         119,
         122, 123, 129, 132, 143, 144, 151},
        {1, 2, 6, 8, 11, 14, 24, 26, 29, 37, 48, 51, 54, 59, 60, 71, 74, 76, 77, 81, 98, 101, 103, 111, 114, 116, 119,
         122,
         123, 131, 133, 137, 139, 140, 151},
        {1, 2, 7, 8, 9, 10, 13, 24, 35, 46, 54, 60, 71, 72, 98, 99, 104, 108, 109, 110, 111, 112, 116, 117, 119, 121,
         122,
         132, 143, 144, 147, 148, 151, 152},
        {1, 3, 4, 5, 6, 11, 14, 25, 26, 27, 29, 37, 54, 60, 71, 74, 76, 77, 81, 85, 86, 98, 101, 104, 111, 114, 116,
         119,
         122, 123, 129, 132, 143, 144, 151},
        {1, 2, 6, 8, 11, 14, 24, 26, 29, 37, 48, 51, 54, 59, 60, 71, 74, 81, 98, 101, 103, 111, 114, 116, 119, 129, 132,
         135, 136, 138, 141, 143, 147, 151},
        {1, 2, 6, 11, 14, 49, 50, 52, 53, 55, 56, 57, 71, 74, 81, 98, 101, 103, 111, 114, 116, 119, 122, 123, 129, 132,
         134,
         135, 136, 138, 143, 144, 151},
        {1, 2, 6, 11, 14, 24, 26, 29, 37, 48, 51, 54, 58, 59, 60, 71, 74, 76, 77, 81, 98, 101, 103, 111, 114, 116, 119,
         122,
         123, 128, 129, 132, 143, 144},
        {1, 2, 6, 11, 14, 24, 26, 29, 37, 49, 50, 52, 53, 55, 56, 57, 71, 74, 76, 77, 81, 98, 101, 103, 111, 116, 119,
         122,
         123, 129, 130, 132, 143, 144},
        {1, 3, 4, 5, 6, 11, 12, 14, 24, 28, 30, 31, 33, 34, 35, 36, 46, 54, 60, 89, 98, 101, 104, 111, 114, 116, 119,
         122,
         123, 124, 129, 132, 143, 144},
        {1, 4, 5, 15, 20, 23, 37, 38, 39, 40, 42, 44, 47, 71, 72, 76, 85, 87, 98, 99, 103, 104, 113, 122, 129, 130, 132,
         143, 144, 147, 148, 151, 152},
        {3, 5, 9, 10, 13, 17, 20, 21, 22, 24, 26, 27, 29, 37, 85, 87, 98, 101, 103, 111, 114, 116, 119, 129, 132, 135,
         136,
         138, 141, 143, 147, 151},
        {1, 4, 15, 20, 23, 37, 38, 39, 40, 42, 43, 44, 71, 72, 76, 85, 87, 98, 99, 103, 104, 113, 122, 129, 130, 132,
         143,
         144, 147, 148, 151, 152},
        {1, 2, 6, 11, 14, 24, 26, 28, 29, 37, 49, 50, 52, 53, 55, 56, 57, 71, 74, 75, 76, 77, 81, 98, 113, 115, 122,
         123,
         129, 130, 132, 143, 144},
        {24, 26, 28, 29, 37, 49, 50, 52, 53, 55, 56, 57, 58, 61, 62, 63, 67, 71, 74, 75, 81, 97, 102, 113, 115, 122,
         123,
         129, 130, 132, 143, 144},
        {24, 26, 28, 29, 37, 49, 50, 52, 53, 55, 56, 57, 58, 61, 62, 63, 67, 70, 71, 74, 75, 81, 98, 113, 115, 122, 123,
         129, 130, 132, 143, 144},
        {1, 2, 6, 7, 8, 9, 10, 13, 24, 35, 46, 54, 60, 71, 72, 98, 99, 104, 108, 109, 110, 111, 112, 116, 117, 119, 121,
         122, 132, 143, 144, 154},
        {1, 2, 7, 8, 9, 10, 13, 24, 35, 46, 54, 60, 71, 72, 76, 81, 82, 98, 99, 103, 105, 106, 107, 113, 122, 132, 143,
         144,
         147, 148, 151, 152},
        {1, 4, 15, 20, 35, 41, 45, 46, 71, 72, 76, 81, 82, 85, 87, 98, 99, 103, 104, 105, 113, 122, 129, 130, 132, 143,
         144,
         147, 148, 151, 152},
        {1, 2, 7, 8, 9, 10, 13, 24, 35, 46, 54, 60, 98, 99, 104, 108, 109, 111, 112, 116, 117, 119, 121, 126, 132, 143,
         144,
         147, 148, 151, 152},
        {14, 16, 17, 18, 19, 22, 37, 39, 40, 42, 43, 44, 47, 71, 72, 76, 85, 87, 98, 99, 103, 104, 113, 122, 129, 130,
         132,
         147, 148, 151, 152},
        {1, 3, 4, 5, 6, 11, 12, 14, 24, 30, 33, 35, 36, 46, 54, 59, 60, 76, 77, 81, 85, 86, 90, 94, 96, 97, 113, 115,
         122,
         123, 132, 143, 144},
        {1, 4, 5, 15, 20, 21, 22, 23, 25, 35, 38, 43, 45, 46, 71, 72, 85, 87, 98, 99, 103, 104, 105, 113, 122, 132, 151,
         152, 154, 157, 158},
        {3, 5, 13, 20, 21, 22, 24, 26, 27, 29, 30, 32, 37, 85, 87, 98, 101, 111, 114, 116, 119, 129, 132, 135, 136, 138,
         141, 143, 147, 151},
        {24, 26, 27, 28, 29, 37, 49, 50, 52, 53, 55, 56, 57, 58, 61, 62, 63, 67, 73, 76, 77, 81, 97, 102, 113, 115, 129,
         130, 132, 143, 144},
        {14, 16, 17, 19, 37, 39, 40, 42, 44, 60, 61, 65, 69, 71, 72, 76, 85, 87, 98, 99, 103, 104, 113, 122, 132, 147,
         148,
         151, 152, 157},
        {2, 7, 8, 9, 10, 13, 24, 28, 30, 31, 33, 35, 36, 46, 88, 89, 98, 101, 104, 111, 114, 116, 119, 122, 123, 124,
         129,
         132, 143, 144},
        {3, 5, 9, 13, 17, 20, 21, 22, 24, 26, 27, 29, 37, 85, 87, 98, 101, 103, 111, 114, 116, 119, 141, 143, 147, 151,
         154,
         155, 157},
        {3, 4, 5, 6, 11, 12, 14, 24, 30, 35, 36, 46, 54, 59, 60, 89, 91, 92, 93, 95, 98, 101, 113, 115, 122, 123, 124,
         132,
         143, 144},
        {1, 2, 6, 7, 8, 9, 10, 13, 24, 35, 46, 54, 60, 71, 72, 98, 99, 103, 105, 106, 107, 113, 120, 122, 132, 143, 144,
         147, 148},
        {1, 2, 6, 8, 11, 12, 14, 24, 26, 29, 37, 71, 74, 81, 98, 101, 103, 111, 114, 116, 119, 122, 123, 127, 131, 137,
         139,
         140},
        {1, 3, 4, 5, 6, 12, 14, 24, 28, 30, 33, 35, 36, 46, 54, 59, 60, 89, 98, 101, 113, 115, 118, 122, 123, 124, 132,
         143,
         144},
        {2, 6, 7, 8, 9, 10, 13, 24, 28, 30, 33, 35, 36, 46, 54, 60, 104, 111, 119, 122, 140, 145, 146, 147, 151, 154,
         157,
         159},
        {3, 4, 5, 6, 11, 12, 14, 24, 35, 46, 54, 60, 89, 93, 95, 98, 99, 100, 113, 115, 118, 122, 123, 124, 132, 143,
         144},
        {5, 20, 21, 22, 23, 28, 35, 43, 46, 85, 87, 111, 112, 119, 122, 125, 141, 143, 145, 146, 147, 151, 154, 157,
         158},
        {5, 20, 28, 35, 46, 85, 87, 111, 112, 119, 129, 132, 133, 135, 136, 138, 143, 144, 147, 148, 149, 151, 152,
         153},
        {5, 20, 24, 28, 33, 35, 36, 43, 46, 85, 87, 111, 119, 122, 129, 132, 135, 138, 141, 143, 150, 154, 157, 159},
        {1, 2, 4, 6, 7, 8, 11, 12, 14, 25, 26, 27, 29, 37, 81, 98, 101, 103, 111, 114, 116, 119, 140, 142, 147, 151},
        {24, 26, 27, 28, 29, 30, 37, 60, 64, 65, 66, 69, 71, 75, 81, 97, 102, 113, 115, 122, 123, 129, 130, 132},
        {6, 11, 14, 24, 25, 26, 27, 28, 29, 32, 37, 48, 54, 55, 58, 59, 60, 78, 85, 86, 97, 113, 132, 143, 144},
        {25, 31, 35, 36, 46, 53, 55, 57, 58, 61, 62, 63, 67, 70, 73, 76, 79, 81, 97, 129, 130, 132, 143, 144},
        {25, 35, 36, 46, 52, 53, 55, 56, 57, 58, 61, 62, 63, 67, 70, 73, 78, 84, 97, 129, 130, 132, 143, 144},
        {25, 35, 36, 46, 53, 55, 57, 58, 61, 62, 63, 67, 70, 73, 76, 79, 80, 81, 97, 129, 130, 132, 143, 144},
        {5, 20, 21, 22, 28, 30, 31, 33, 35, 36, 43, 46, 85, 87, 111, 119, 122, 141, 143, 154, 155, 156, 157},
        {24, 26, 28, 29, 30, 37, 60, 64, 65, 66, 68, 69, 71, 74, 75, 81, 98, 101, 113, 115, 122, 123, 157},
        {3, 4, 5, 6, 11, 12, 14, 24, 35, 46, 54, 59, 60, 71, 72, 78, 83, 85, 86, 97, 113, 132, 143, 144},

    ],
}


# ===  ===
def get_bounds():
    mins = np.array([EXPERIMENT_CONFIG['MIN_X'], EXPERIMENT_CONFIG['MIN_Y'], EXPERIMENT_CONFIG['MIN_Z']])
    maxs = np.array([EXPERIMENT_CONFIG['MAX_X'], EXPERIMENT_CONFIG['MAX_Y'], EXPERIMENT_CONFIG['MAX_Z']])
    return mins, maxs

def clip_state(state):
    mins, maxs = get_bounds()
    return np.clip(state, mins, maxs)

def denormalize_state(normalized_state):
    """"""
    mins, maxs = get_bounds()
    return normalized_state * (maxs - mins) / 2 + (mins + maxs) / 2


def coverage_similarity(triggered, target_path):
    """
    Similarity: / target paths
    """
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0

    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)


def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']

    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']

    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']

    return reward


# ===   ===
def execute_Tr(a):
    """
    替换原有的编排规则函数
    参数a: 包含3个元素的元组或数组，分别对应path_depth, file_count, access_level
    返回: 触发的规则编号集合
    """
    path_depth, file_count, access_level = float(a[0]), int(a[1]), float(a[2])

    # 使用正确的变量名
    x, y, z = path_depth, file_count, access_level

    triggered = set()
    # 创建一个字典来存储b数组的值，用于跟踪哪些规则被触发
    b = {}
    x_score = max(0, 100 - x * 1.67)

    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[0] = 1
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((z * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[1] = 2
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * x) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[2] = 3
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[3] = 4
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * 43) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[4] = 5
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((74 * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[5] = 6
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (47 + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[6] = 7
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (x + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[7] = 8
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (y + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[8] = 9
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 10) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[9] = 10
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) != 50 and x ** 2 + y ** 2 > z ** 2):
        b[10] = 11
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 29 and x ** 2 + y ** 2 > z ** 2):
        b[11] = 12
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 69 and x ** 2 + y ** 2 > z ** 2):
        b[12] = 13
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2):
        b[13] = 14
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and y ** 2 + y ** 2 > z ** 2):
        b[14] = 15
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 3 + y ** 2 > z ** 2):
        b[15] = 16
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + x ** 2 > z ** 2):
        b[16] = 17
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 3 > z ** 2):
        b[17] = 18
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 != z ** 2):
        b[18] = 19
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2.5):
        b[19] = 20
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 1.4 > z ** 2):
        b[20] = 21
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + 34 ** 2 > z ** 2):
        b[21] = 22
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and 64 ** 2 + y ** 2 > z ** 2):
        b[22] = 23
    if (x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2:
        pattern_type = "311A1"
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((z ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[23] = 24
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - z ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[24] = 25
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 1.5) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[25] = 26
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2.3 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[26] = 27
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((45 ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[27] = 28
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - 31 ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[28] = 29
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[29] = 30
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (x + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[30] = 31
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 2) < -30 or (abs(x - y) * z) / 100 > 45):
        b[31] = 32
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (26 + 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[32] = 33
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y - 0.1) < -30 or (abs(x - y) * z) / 100 > 45):
        b[33] = 34
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) != -30 or (abs(x - y) * z) / 100 > 45):
        b[34] = 35
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -10 or (abs(x - y) * z) / 100 > 45):
        b[35] = 36
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 and (abs(x - y) * z) / 100 > 45):
        b[36] = 37
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(z - y) * z) / 100 > 45):
        b[37] = 38
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - z) * z) / 100 > 45):
        b[38] = 39
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * y) / 100 > 45):
        b[39] = 40
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * x) / 100 > 45):
        b[40] = 41
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * 34) / 100 > 45):
        b[41] = 42
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - 26) * z) / 100 > 45):
        b[42] = 43
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(52 - y) * z) / 100 > 45):
        b[43] = 44
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 76 > 45):
        b[44] = 45
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 != 45):
        b[45] = 46
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 55):
        b[46] = 47
    if (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45:
        pattern_type = "311A2"

    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + x) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35):
        b[47] = 48
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((y + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35):
        b[48] = 49
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((24 + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35):
        b[49] = 50
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + 31) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35):
        b[50] = 51
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 1.5) ** 2 < z * 20 and (x * x * z) / 1000 < 35):
        b[51] = 52
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2.4 < z * 20 and (x * x * z) / 1000 < 35):
        b[52] = 53
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 != z * 20 and (x * x * z) / 1000 < 35):
        b[53] = 54
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < 25 * 20 and (x * x * z) / 1000 < 35):
        b[54] = 55
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 14 and (x * x * z) / 1000 < 35):
        b[55] = 56
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < x * 20 and (x * x * z) / 1000 < 35):
        b[56] = 57
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < y * 20 and (x * x * z) / 1000 < 35):
        b[57] = 58
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 1.2 < z * 20 and (x * x * z) / 1000 < 35):
        b[58] = 59
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 or (x * x * z) / 1000 < 35):
        b[59] = 60
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (y * x * z) / 1000 < 35):
        b[60] = 61
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (z * x * z) / 1000 < 35):
        b[61] = 62
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * 35 * z) / 1000 < 35):
        b[62] = 63
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * x * x) / 1000 < 35):
        b[63] = 64
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * x * y) / 1000 < 35):
        b[64] = 65
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * x * 26) / 1000 < 35):
        b[65] = 66
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * 54 * z) / 1000 < 35):
        b[66] = 67
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1050 < 35):
        b[67] = 68
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 != 35):
        b[68] = 69
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 25):
        b[69] = 70
    if ((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35:
        pattern_type = "311A3"
    if (x < 35 and y < 50) != (x < 35 or y < 50):
        b[70] = 71
    if (x < 35 and y < 50) != (x != 35 and y < 50):
        b[71] = 72
    if (x < 35 and y < 50) != (x < 19 and y < 50):
        b[72] = 73
    if (x < 35 and y < 50) != (x < 35 and y != 50):
        b[73] = 74
    if (x < 35 and y < 50) != (x < 35 and y < 61):
        b[74] = 75
    if x < 35 and y < 50:
        pattern_type = 312
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y < 35 or z < 65 and x < 50)):
        b[75] = 76
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y != 35 and z < 65 and x < 50)):
        b[76] = 77
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y < 25 and z < 65 and x < 50)):
        b[77] = 78
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y < 35 and z != 65 and x < 50)):
        b[78] = 79
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y < 35 and z < 72 and x < 50)):
        b[79] = 80
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y < 35 and z < 65 or x < 50)):
        b[80] = 81
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y < 35 and z < 65 and x != 50)):
        b[81] = 82
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 31)):
        b[82] = 83
    if ((y < 30 and z < 60) or (y < 35 and z < 65 and x < 50)) != ((y < 30 and z < 60) or (y < 35 and z < 53 and x < 50)):
        b[83] = 84
    if (y < 30 and z < 60) or (y < 35 and z < 65 and x < 50):
        pattern_type = 313
    if (z < 40 and x > 50) != (z < 40 or x > 50):
        b[84] = 85
    if (z < 40 and x > 50) != (z < 40 and x != 50):
        b[85] = 86
    if (z < 40 and x > 50) != (z != 40 and x > 50):
        b[86] = 87
    if (z < 40 and x > 50) != (z < 29 and x > 50):
        b[87] = 88
    if (z < 40 and x > 50) != (z < 40 and x > 62):
        b[88] = 89
    if z < 40 and x > 50:
        pattern_type = 314
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (y * y) / 100 < 28) or (x < 60 and y < 55)):
        b[89] = 90
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * x) / 100 < 28) or (x < 60 and y < 55)):
        b[90] = 91
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * 53) / 100 < 28) or (x < 60 and y < 55)):
        b[91] = 92
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (41 * y) / 100 < 28) or (x < 60 and y < 55)):
        b[92] = 93
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 88 < 28) or (x < 60 and y < 55)):
        b[93] = 94
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 100 != 28) or (x < 60 and y < 55)):
        b[94] = 95
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 100 < 21) or (x < 60 and y < 55)):
        b[95] = 96
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) and (x < 60 and y < 55)):
        b[96] = 97
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 or y < 55)):
        b[97] = 98
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x != 60 and y < 55)):
        b[98] = 99
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 69 and y < 55)):
        b[99] = 100
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y != 55)):
        b[100] = 101
    if ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55)) != ((40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 45)):
        b[101] = 102
    if (40 <= x < 65 and y < 60 and (x * y) / 100 < 28) or (x < 60 and y < 55):
        pattern_type = 315
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((x ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)):
        b[102] = 103
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((z ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)):
        b[103] = 104
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2.4) / (x + z + 1) < 22) or (y < 65 and x < 70)):
        b[104] = 105
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (z + z + 1) < 22) or (y < 65 and x < 70)):
        b[105] = 106
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (y + z + 1) < 22) or (y < 65 and x < 70)):
        b[106] = 107
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + x + 1) < 22) or (y < 65 and x < 70)):
        b[107] = 108
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + y + 1) < 22) or (y < 65 and x < 70)):
        b[108] = 109
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 7) < 22) or (y < 65 and x < 70)):
        b[109] = 110
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 1) != 22) or (y < 65 and x < 70)):
        b[110] = 111
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 1) < 32) or (y < 65 and x < 70)):
        b[111] = 112
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 1) < 22) and (y < 65 and x < 70)):
        b[112] = 113
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 1) < 22) or (y != 65 and x < 70)):
        b[113] = 114
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 1) < 22) or (y < 45 and x < 70)):
        b[114] = 115
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 1) < 22) or (y < 65 or x < 70)):
        b[115] = 116
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x != 70)):
        b[116] = 117
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 55)):
        b[117] = 118
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((24 ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)):
        b[118] = 119
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (56 + z + 1) < 22) or (y < 65 and x < 70)):
        b[119] = 120
    if (((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70)) != (((y ** 2) / (x + 73 + 1) < 22) or (y < 65 and x < 70)):
        b[120] = 121
    if ((y ** 2) / (x + z + 1) < 22) or (y < 65 and x < 70):
        pattern_type = 316
    if (50 <= y < 75 and x > 65) != (50 <= y < 75 or x > 65):
        b[121] = 122
    if (50 <= y < 75 and x > 65) != (50 <= y < 75 and x != 65):
        b[122] = 123
    if (50 <= y < 75 and x > 65) != (50 <= y < 75 and x > 47):
        b[123] = 124
    if (50 <= y < 75 and x > 65) != (50 <= y < 82 and x > 65):
        b[124] = 125
    if (50 <= y < 75 and x > 65) != (57 <= y < 75 and x > 65):
        b[125] = 126
    if 50 <= y < 75 and x > 65:
        pattern_type = 317
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 89 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)):
        b[126] = 127
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((41 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)):
        b[127] = 128
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 or y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)):
        b[128] = 129
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y != 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)):
        b[129] = 130
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 79 and (z * y) / 100 < 52) or (z < 75 and y > 75)):
        b[130] = 131
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 or (z * y) / 100 < 52) or (z < 75 and y > 75)):
        b[131] = 132
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (y * y) / 100 < 52) or (z < 75 and y > 75)):
        b[132] = 133
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * z) / 100 < 52) or (z < 75 and y > 75)):
        b[133] = 134
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * 26) / 100 < 52) or (z < 75 and y > 75)):
        b[134] = 135
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (53 * y) / 100 < 52) or (z < 75 and y > 75)):
        b[135] = 136
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * y) / 89 < 52) or (z < 75 and y > 75)):
        b[136] = 137
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * y) / 100 != 52) or (z < 75 and y > 75)):
        b[137] = 138
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * y) / 100 < 39) or (z < 75 and y > 75)):
        b[138] = 139
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) and (z < 75 and y > 75)):
        b[139] = 140
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z != 75 and y > 75)):
        b[140] = 141
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 65 and y > 75)):
        b[141] = 142
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 or y > 75)):
        b[142] = 143
    if ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75)) != ((55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y != 75)):
        b[143] = 144
    if (55 <= z < 80 and y > 70 and (z * y) / 100 < 52) or (z < 75 and y > 75):
        pattern_type = 318
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x != 88 and y > 75) or (78 <= x < 90 and y > 72)):
        b[144] = 145
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x < 94 and y > 75) or (78 <= x < 90 and y > 72)):
        b[145] = 146
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x < 88 or y > 75) or (78 <= x < 90 and y > 72)):
        b[146] = 147
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x < 88 and y != 75) or (78 <= x < 90 and y > 72)):
        b[147] = 148
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x < 88 and y > 67) or (78 <= x < 90 and y > 72)):
        b[148] = 149
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x < 88 and y > 75) and (78 <= x < 90 and y > 72)):
        b[149] = 150
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x < 88 and y > 75) or (78 <= x < 90 or y > 72)):
        b[150] = 151
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y != 72)):
        b[151] = 152
    if ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72)) != ((75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 62)):
        b[152] = 153
    if (75 <= x < 88 and y > 75) or (78 <= x < 90 and y > 72):
        pattern_type = 319
    if (x >= 88 and y >= 85 and z >= 85) != (x >= 88 or y >= 85 and z >= 85):
        b[153] = 154
    if (x >= 88 and y >= 85 and z >= 85) != (x != 88 and y >= 85 and z >= 85):
        b[154] = 155
    if (x >= 88 and y >= 85 and z >= 85) != (x >= 78 and y >= 85 and z >= 85):
        b[155] = 156
    if (x >= 88 and y >= 85 and z >= 85) != (x >= 88 and y >= 85 or z >= 85):
        b[156] = 157
    if (x >= 88 and y >= 85 and z >= 85) != (x >= 88 and y != 85 and z >= 85):
        b[157] = 158
    if (x >= 88 and y >= 85 and z >= 85) != (x >= 88 and y >= 85 and z != 85):
        b[158] = 159
    if x >= 88 and y >= 85 and z >= 85:
        pattern_type = 320

    # 返回被触发的规则编号集合
    return set(b.values())

# 
execute_Tr = execute_Tr

# === DQN ===
class DQNNetwork(nn.Module):
    def __init__(self, action_size=30):
        super(DQNNetwork, self).__init__()
        hidden_dim = EXPERIMENT_CONFIG['HIDDEN_DIM']

        self.fc1 = nn.Linear(3, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)
        self.output = nn.Linear(hidden_dim, action_size)

    def forward(self, x):
        x = torch.relu(self.fc1(x))
        x = torch.relu(self.fc2(x))
        x = torch.relu(self.fc3(x))
        return self.output(x)


# === Path  ===
class PathReplayBuffer:
    """Path """

    def __init__(self, path_idx, capacity=20000):
        self.path_idx = path_idx
        self.capacity = capacity
        self.buffer = deque(maxlen=capacity)
        self.similarities = deque(maxlen=capacity)  # Similarity

    def push(self, state, action, reward, next_state, done, similarity):
        self.buffer.append((state, action, reward, next_state, done))
        self.similarities.append(similarity)

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return None

        batch = random.sample(self.buffer, batch_size)
        state, action, reward, next_state, done = map(np.stack, zip(*batch))

        return (
            torch.FloatTensor(state).to(device),
            torch.LongTensor(action).to(device),
            torch.FloatTensor(reward).to(device),
            torch.FloatTensor(next_state).to(device),
            torch.BoolTensor(done).to(device)
        )

    def get_top_k(self, k=20):
        """Path Top-K"""
        if len(self.buffer) == 0:
            return []

        # buffersimilarities
        samples_with_sim = list(zip(self.buffer, self.similarities))
        samples_with_sim.sort(key=lambda x: x[1], reverse=True)

        # Top-K
        top_k = samples_with_sim[:k]

        results = []
        for (state, _, _, _, _), similarity in top_k:
            original_state = denormalize_state(state)
            original_state_int = np.round(original_state).astype(int)

            results.append({
                'state': original_state_int,
                'similarity': similarity,
                'triggered': execute_Tr(*original_state_int)
            })

        return results

    def __len__(self):
        return len(self.buffer)


# === DQN(:)===
class ImprovedDQNAgent:
    def __init__(self, num_paths, action_size=30):
        self.action_size = action_size
        self.num_paths = num_paths
        self.epsilon = 0.9
        self.epsilon_min = 0.1
        self.epsilon_decay = 0.995

        self.q_network = DQNNetwork(action_size).to(device)
        self.target_network = DQNNetwork(action_size).to(device)

        lr = EXPERIMENT_CONFIG['LEARNING_RATE']
        self.optimizer = optim.Adam(self.q_network.parameters(), lr=lr)

        # Path 
        capacity = EXPERIMENT_CONFIG['REPLAY_BUFFER_CAPACITY']
        self.replay_buffers = {}
        for path_idx in range(num_paths):
            self.replay_buffers[path_idx] = PathReplayBuffer(path_idx, capacity)

        self.replay_train_count = 0
        self.update_target_network()

    def discrete_to_action_delta(self, action_idx):
        # 
        delta_values = [5, 3, 2, 1, 0.5, -0.5, -1, -2, -3, -5]

        if action_idx >= 30:
            action_idx = action_idx % 30

        dim = action_idx // 10
        delta_idx = action_idx % 10
        delta = delta_values[delta_idx]

        action_delta = np.zeros(3)
        action_delta[dim] = delta

        return action_delta

    def get_action(self, state):
        mins, maxs = get_bounds()
        normalized_state = (state - (mins + maxs) / 2) / ((maxs - mins) / 2)

        if random.random() < self.epsilon:
            action_idx = random.randint(0, self.action_size - 1)
        else:
            state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)
            with torch.no_grad():
                q_values = self.q_network(state_tensor)
                action_idx = q_values.argmax().item()

        action_delta = self.discrete_to_action_delta(action_idx)
        return action_delta, action_idx

    def store_experience(self, path_idx, state, action_idx, reward, next_state, done, similarity):
        """Path """
        mins, maxs = get_bounds()
        normalized_state = (state - (mins + maxs) / 2) / ((maxs - mins) / 2)
        normalized_next_state = (next_state - (mins + maxs) / 2) / ((maxs - mins) / 2)

        self.replay_buffers[path_idx].push(
            normalized_state, action_idx, reward,
            normalized_next_state, done, similarity
        )

    def replay_train(self, path_idx):
        """Path """
        batch_size = EXPERIMENT_CONFIG['REPLAY_BATCH_SIZE']
        batch = self.replay_buffers[path_idx].sample(batch_size)

        if batch is None:
            return

        states, actions, rewards, next_states, dones = batch

        current_q_values = self.q_network(states).gather(1, actions.unsqueeze(1))

        with torch.no_grad():
            next_q_values = self.target_network(next_states).max(1)[0]
            target_q_values = rewards + (0.99 * next_q_values * ~dones)

        loss = F.mse_loss(current_q_values.squeeze(), target_q_values)

        self.optimizer.zero_grad()
        loss.backward()
        torch.nn.utils.clip_grad_norm_(self.q_network.parameters(), 1.0)
        self.optimizer.step()

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

        self.replay_train_count += 1

        if self.replay_train_count % 2 == 0:
            self.update_target_network()
            print(f"    ->  (Run {self.replay_train_count})")

    def update_target_network(self):
        self.target_network.load_state_dict(self.q_network.state_dict())

    def get_all_top_k(self, k=20):
        """Path Top-K"""
        results = {}
        for path_idx in range(self.num_paths):
            results[path_idx] = self.replay_buffers[path_idx].get_top_k(k)
        return results

    def get_buffer_stats(self):
        """"""
        stats = {}
        for path_idx in range(self.num_paths):
            stats[path_idx] = len(self.replay_buffers[path_idx])
        return stats


# === Metric ===
def calculate_run_performance(run_idx, dqn_results, training_time, total_steps, update_count, agent):
    """ runMetric"""
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    # 1. (Total Reward)
    total_reward = 0
    # 2. (Average Reward)
    average_reward = 0
    # 5. (Convergence)
    convergence = 0
    # 12. (Environment Adaptability)
    environment_adaptability = 0
    # 13. (Generalization Ability)
    generalization_ability = 0
    # 15. (Computational Efficiency)
    computational_efficiency = 0
    # 16. (Policy Update Frequency)
    policy_update_frequency = 0

    # Similarity
    all_similarities = []

    # Metric
    total_samples = 0
    all_rewards = []

    for path_idx in range(num_paths):
        samples = dqn_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']

            total_reward += reward
            all_rewards.append(reward)
            all_similarities.append(similarity)
            total_samples += 1

    # 1. 
    total_reward = total_reward

    # 2. 
    if total_samples > 0:
        average_reward = total_reward / total_samples

    # 5. (Average Similarity)
    if all_similarities:
        convergence = np.mean(all_similarities)

    # 12. (Similarity)
    if len(all_similarities) > 1:
        environment_adaptability = 1 / (np.std(all_similarities) + 1e-8)

    # 13. (Average Similarity)
    generalization_ability = convergence

    # 15. (/ seconds)
    if training_time > 0:
        computational_efficiency = total_steps / training_time

    # 16. 
    if training_time > 0:
        policy_update_frequency = update_count / training_time

    # Similarity
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0

    return {
        'Run': run_idx + 1,

        # Metric
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),

        # Similarity
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }


# === Excel ===
def export_to_excel(all_dqn_results, all_performance_data, target_paths, output_path="DQN_20 run.xlsx"):
    """20 runDQNExcel"""
    print("\nExcel...")

    # 
    all_dqn_summary_data = []
    all_dqn_detailed_data = []

    #  run
    for run_idx, (dqn_results, performance_data) in enumerate(zip(all_dqn_results, all_performance_data)):
        # ===== Sheet1: DQNPath  =====
        dqn_summary_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = dqn_results[path_idx]

            if len(samples) == 0:
                dqn_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue

            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = 'Yes' if perfect_count > 0 else 'No'

            dqn_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })
        all_dqn_summary_data.extend(dqn_summary_data)

        # ===== Sheet2: DQNDetailed Sample Data =====
        dqn_detailed_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = dqn_results[path_idx]

            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']

                dqn_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Sample ID': sample_idx + 1,
                    'X': int(state[0]),
                    'Y': int(state[1]),
                    'Z': int(state[2]),
                    'Similarity': round(similarity, 4),
                    'Perfect Coverage': 'Yes' if abs(similarity - 1.0) < 0.001 else 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path))),
                    'Triggered Rules': ', '.join(map(str, sorted(triggered))),
                    'Hit Rule Count': len(target_path.intersection(triggered)),
                    'Target Rule Count': len(target_path)
                })
        all_dqn_detailed_data.extend(dqn_detailed_data)

    # Excel
    dqn_summary_df = pd.DataFrame(all_dqn_summary_data)
    dqn_detailed_df = pd.DataFrame(all_dqn_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet1: DQNPath 
        dqn_summary_df.to_excel(writer, sheet_name='DQNPath ', index=False)

        # Sheet2: DQNDetailed Sample Data
        dqn_detailed_df.to_excel(writer, sheet_name='DQNDetailed Sample Data', index=False)

        # Sheet3: Metric
        selected_columns = [
            'Run',
            'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df_selected = performance_df[selected_columns]
        performance_df_selected.to_excel(writer, sheet_name='Metric', index=False)

        # 
        workbook = writer.book

        # 
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        perfect_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # 

        # === Sheet1 ===
        ws1 = writer.sheets['DQNPath ']
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 
        for row_idx in range(2, ws1.max_row + 1):
            if ws1.cell(row_idx, 9).value == '':  # Run 9""
                for col_idx in range(1, ws1.max_column + 1):
                    ws1.cell(row_idx, col_idx).fill = perfect_fill

        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 15
        ws1.column_dimensions['F'].width = 15
        ws1.column_dimensions['G'].width = 15
        ws1.column_dimensions['H'].width = 15
        ws1.column_dimensions['I'].width = 15
        ws1.column_dimensions['J'].width = 50

        # === Sheet2 ===
        ws2 = writer.sheets['DQNDetailed Sample Data']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 15
        ws2.column_dimensions['I'].width = 40
        ws2.column_dimensions['J'].width = 40
        ws2.column_dimensions['K'].width = 15
        ws2.column_dimensions['L'].width = 15

        # === Sheet3 ===
        ws3 = writer.sheets['Metric']
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        for col in columns:
            ws3.column_dimensions[col].width = 18

    print(f"Excel: {output_path}")
    print(f"  - Sheet1: DQNPath  ({len(all_dqn_summary_data)})")
    print(f"  - Sheet2: DQNDetailed Sample Data ({len(all_dqn_detailed_data)})")
    print(f"  - Sheet3: Metric ({len(all_performance_data)})")


# === DQN training(:)===
def train_dqn_workflow():
    print("=" * 80)
    print("DQN training ()")
    print("Similarity:  / target paths")
    print(
        f": Path {EXPERIMENT_CONFIG['NUM_ROUNDS']},{EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(f": Path ,={EXPERIMENT_CONFIG['REPLAY_BUFFER_CAPACITY']}")
    print("=" * 80)

    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    # (Number of Paths)
    agent = ImprovedDQNAgent(num_paths=num_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n: Path {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']}")
    path_samples = {}
    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            state = np.array([
                random.randint(EXPERIMENT_CONFIG['MIN_X'], EXPERIMENT_CONFIG['MAX_X']),
                random.randint(EXPERIMENT_CONFIG['MIN_Y'], EXPERIMENT_CONFIG['MAX_Y']),
                random.randint(EXPERIMENT_CONFIG['MIN_Z'], EXPERIMENT_CONFIG['MAX_Z'])
            ], dtype=np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  Path  {path_idx + 1}/{num_paths}:  {len(samples)} ")

    # 
    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size
    num_rounds = EXPERIMENT_CONFIG['NUM_ROUNDS']

    print(f"\n:")
    print(f"  - : {batch_size}")
    print(f"  - Path : {num_batches}")
    print(f"  - Path : {num_rounds}")
    print(f"  - : {EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(
        f"  - : {num_paths} Path  x {num_rounds}  x {num_batches}  = {num_paths * num_rounds * num_batches} ")
    print("-" * 80)

    # :completedPath ,Path 
    for path_idx in range(num_paths):
        target_path = target_paths[path_idx]
        print(f"\n{'=' * 80}")
        print(f"Start training path  {path_idx + 1}/{num_paths}")
        print(f": {sorted(target_path)}")
        print(f": replay_buffers[{path_idx}]")
        print(f"{'=' * 80}")

        # Path NUM_ROUNDS
        for round_idx in range(num_rounds):
            print(f"\n{'' * 80}")
            print(f"Path  {path_idx + 1} - Run  {round_idx + 1}/{num_rounds} ")
            print(f"{'' * 80}")

            # Per roundnum_batches
            for batch_idx in range(num_batches):
                print(f"\n   {batch_idx + 1}/{num_batches} (Path {path_idx + 1}, Run {round_idx + 1})")

                # 
                batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]

                batch_rewards = []
                batch_similarities = []

                # 
                for sample_idx, initial_state in enumerate(batch_samples):
                    state = initial_state.copy()
                    episode_reward = 0
                    final_similarity = 0

                    # STEPS_PER_SAMPLE
                    for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                        action_delta, action_idx = agent.get_action(state)

                        next_state = state + action_delta
                        next_state = clip_state(next_state)

                        triggered = execute_Tr(*next_state)  #
                        reward = unified_reward_function(triggered, target_path)
                        similarity = coverage_similarity(triggered, target_path)

                        done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)

                        # Path 
                        agent.store_experience(
                            path_idx, state, action_idx, reward, next_state, done, similarity
                        )

                        state = next_state
                        episode_reward += reward
                        final_similarity = similarity
                        total_steps += 1

                    batch_rewards.append(episode_reward)
                    batch_similarities.append(final_similarity)

                # 
                avg_reward = np.mean(batch_rewards)
                avg_similarity = np.mean(batch_similarities)
                max_similarity = np.max(batch_similarities)
                print(f"    ={avg_reward:.2f}, Average Similarity={avg_similarity:.4f}, "
                      f"Similarity={max_similarity:.4f}, epsilon={agent.epsilon:.3f}")

                # Path 
                print(f"    (Path {path_idx})...")
                agent.replay_train(path_idx)

                # Path 
                buffer_size = len(agent.replay_buffers[path_idx])
                print(f"    Path {path_idx}: {buffer_size}, : {agent.replay_train_count}")

    training_time = time.time() - start_time

    print("\n" + "=" * 80)
    print(f"DQN trainingcompleted! Total elapsed time: {training_time:.2f} seconds, : {total_steps}")
    print(f": {agent.replay_train_count}")
    print(f": {agent.replay_train_count // 2}")

    # Path 
    print("\nPath :")
    buffer_stats = agent.get_buffer_stats()
    for path_idx, size in buffer_stats.items():
        print(f"  Path {path_idx + 1}: {size} ")

    print("=" * 80)

    # Top-K
    print(f"\nPath SimilarityMaximum{EXPERIMENT_CONFIG['TOP_K_SAMPLES']}...")
    dqn_top_k_results = agent.get_all_top_k(EXPERIMENT_CONFIG['TOP_K_SAMPLES'])

    return agent, dqn_top_k_results, training_time, total_steps, agent.replay_train_count


# ===  ===
def main():
    print("\n" + "=" * 80)
    print("DQN - 20 run")
    print("Metric")
    print("=" * 80)

    all_dqn_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    # 20
    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'='*80}")
        print(f"Start run  {run_idx + 1}/{EXPERIMENT_CONFIG['NUM_RUNS']}  run")
        print(f"{'='*80}")

        # DQN training
        dqn_agent, dqn_results, training_time, total_steps, update_count = train_dqn_workflow()

        # Metric
        performance_data = calculate_run_performance(
            run_idx, dqn_results, training_time, total_steps, update_count, dqn_agent
        )

        # 
        all_dqn_results.append(dqn_results)
        all_performance_data.append(performance_data)

        print(f"\nRun {run_idx + 1} completed!")
        print(f"  Total Reward: {performance_data['Total Reward']}")
        print(f"  Average Reward: {performance_data['Average Reward']}")
        print(f"  Convergence: {performance_data['Convergence']}")

    # Excel(20 run)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"DQN_20 run_{timestamp}.xlsx"
    export_to_excel(all_dqn_results, all_performance_data, target_paths, output_path)

    # 
    print("\n" + "=" * 80)
    print("20 run")
    print("=" * 80)

    # Metric
    # Metric Extraction
    total_rewards = [p['Total Reward'] for p in all_performance_data]
    average_rewards = [p['Average Reward'] for p in all_performance_data]
    convergences = [p['Convergence'] for p in all_performance_data]
    environment_adaptabilities = [p['Environment Adaptability'] for p in all_performance_data]
    generalization_abilities = [p['Generalization Ability'] for p in all_performance_data]
    computational_efficiencies = [p['Computational Efficiency'] for p in all_performance_data]
    policy_update_frequencies = [p['Policy Update Frequency'] for p in all_performance_data]
    avg_similarities = [p['Average Similarity'] for p in all_performance_data]

    print(f":")
    print(f"  : {np.mean(total_rewards):.2f}")
    print(f"  Standard deviation: {np.std(total_rewards):.2f}")

    print(f"\n:")
    print(f"  : {np.mean(average_rewards):.4f}")
    print(f"  Standard deviation: {np.std(average_rewards):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(convergences):.4f}")
    print(f"  Standard deviation: {np.std(convergences):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(environment_adaptabilities):.4f}")
    print(f"  Standard deviation: {np.std(environment_adaptabilities):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(generalization_abilities):.4f}")
    print(f"  Standard deviation: {np.std(generalization_abilities):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(computational_efficiencies):.2f}")
    print(f"  Standard deviation: {np.std(computational_efficiencies):.2f}")

    print(f"\n:")
    print(f"  : {np.mean(policy_update_frequencies):.4f}")
    print(f"  Standard deviation: {np.std(policy_update_frequencies):.4f}")

    print(f"\nAverage similarity statistics:")
    print(f"  : {np.mean(avg_similarities):.4f}")
    print(f"  Standard deviation: {np.std(avg_similarities):.4f}")

    print("\n" + "=" * 80)
    print(f" {EXPERIMENT_CONFIG['NUM_RUNS']} completed!")
    print("=" * 80)


if __name__ == "__main__":
    main()