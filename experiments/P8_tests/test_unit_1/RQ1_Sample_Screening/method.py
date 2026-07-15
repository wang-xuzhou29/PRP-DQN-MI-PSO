import os
import random
from collections import deque
import numpy as np
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"


STATE_MIN_X, STATE_MAX_X = 2, 100
STATE_MIN_Y, STATE_MAX_Y = 2, 100
STATE_MIN_Z, STATE_MAX_Z = 2, 100
def generate_input():
    return [
        random.randint(STATE_MIN_X, STATE_MAX_X),
        random.randint(STATE_MIN_Y, STATE_MAX_Y),
        random.randint(STATE_MIN_Z, STATE_MAX_Z)
    ]



def execute_Tr(a):
    """
    替换原有的编排规则函数
    参数a: 包含3个元素的元组或数组，分别对应path_depth, file_count, access_level
    返回: 触发的规则编号集合
    """
    path_depth, file_count, access_level = float(a[0]), int(a[1]), float(a[2])

    # 使用正确的变量名
    x, y, z = path_depth, file_count, access_level

    triggered = set()
    # 创建一个字典来存储b数组的值，用于跟踪哪些规则被触发
    b = {}
    x_score = max(0, 100 - x * 1.67)

    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2): b[
        0] = 1
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((27 * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[1] = 2
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * 63) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[2] = 3
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (36 + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[3] = 4
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 7) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[4] = 5
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 12) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[5] = 6
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[6] = 7
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((z * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[7] = 8
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * x) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[8] = 9
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[9] = 10
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (x + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[10] = 11
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (y + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[11] = 12
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) != 50 and x ** 2 + y ** 2 > z ** 2):
        b[12] = 13
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 70 and x ** 2 + y ** 2 > z ** 2):
        b[13] = 14
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 28 and x ** 2 + y ** 2 > z ** 2):
        b[14] = 15
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and y ** 2 + y ** 2 > z ** 2):
        b[15] = 16
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 1 + y ** 2 > z ** 2):
        b[16] = 17
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 1 > z ** 2):
        b[17] = 18
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + x ** 2 > z ** 2):
        b[18] = 19
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 3):
        b[19] = 20
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x * 2 + y ** 2 > z ** 2): b[
        20] = 21
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2.5):
        b[21] = 22
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z * 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[22] = 23
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z - 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[23] = 24
    if (x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2:
        pattern_type = "291A1"
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 3 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[24] = 25
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 4 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[25] = 26
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - z ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[26] = 27
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 1) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[27] = 28
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[28] = 29
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 1 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[29] = 30
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (x + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[30] = 31
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[31] = 32
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 1.5) < -30 or (abs(x - y) * z) / 100 > 45): b[32] = 33
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z - 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[33] = 34
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) != -30 or (abs(x - y) * z) / 100 > 45): b[34] = 35
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -10 or (abs(x - y) * z) / 100 > 45): b[35] = 36
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 and (abs(x - y) * z) / 100 > 45): b[36] = 37
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (24 ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[37] = 38
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - 19 ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[38] = 39
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (42 + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[39] = 40
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45): b[40] = 41
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - z) * z) / 100 > 45): b[41] = 42
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(52 - y) * z) / 100 > 45): b[42] = 43
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - 42) * z) / 100 > 45): b[43] = 44
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * 55) / 100 > 45): b[44] = 45
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 75 > 45): b[45] = 46
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 25): b[46] = 47
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 65): b[47] = 48
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 != 45): b[48] = 49
    if (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45:
        pattern_type = "291A2"
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((46 + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35): b[49] = 50
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + 37) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35): b[50] = 51
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 3) ** 2 < z * 20 and (x * y * z) / 1000 < 35): b[51] = 52
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 3 < z * 20 and (x * y * z) / 1000 < 35): b[52] = 53
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 != z * 20 and (x * y * z) / 1000 < 35): b[53] = 54
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < x * 20 and (x * y * z) / 1000 < 35): b[54] = 55
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35): b[55] = 56
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 30 and (x * y * z) / 1000 < 35): b[56] = 57
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 or (x * y * z) / 1000 < 35): b[57] = 58
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < 52 * 20 and (x * y * z) / 1000 < 35): b[58] = 59
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 3 < z * 20 and (x * y * z) / 1000 < 35): b[59] = 60
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (y * y * z) / 1000 < 35): b[60] = 61
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (z * y * z) / 1000 < 35): b[61] = 62
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35): b[62] = 63
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * z * z) / 1000 < 35): b[63] = 64
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35): b[64] = 65
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * y * y) / 1000 < 35): b[65] = 66
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1200 < 35): b[66] = 67
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 700 < 35): b[67] = 68
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 15): b[68] = 69
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 55): b[69] = 70
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 != 35): b[70] = 71
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (34 * y * z) / 1000 < 35): b[71] = 72
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * 53 * z) / 1000 < 35): b[72] = 73
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * y * 72) / 1000 < 35): b[73] = 74
    if ((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35:
        pattern_type = "291A3"
    if (x_score < 30 and z < 50) != (x_score < 30 or z < 50): b[74] = 75
    if (x_score < 30 and z < 50) != (x_score < 30 and z < 40): b[75] = 76
    if (x_score < 30 and z < 50) != (x_score < 30 and z != 50): b[76] = 77
    if (x_score < 30 and z < 50) != (x_score != 30 and z < 50): b[77] = 78
    if (x_score < 30 and z < 50) != (x_score < 19 and z < 50): b[78] = 79
    if x_score < 30 and z < 50:
        pattern_type = 292
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 35 or y < 60) or (z < 40 and y < 65 and x > 40)): b[79] = 80
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 35 and y < 60) and (z < 40 and y < 65 and x > 40)): b[80] = 81
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 35 and y < 60) or (z < 40 or y < 65 and x > 40)): b[81] = 82
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 35 and y < 60) or (z < 40 and y < 65 or x > 40)): b[82] = 83
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 35 and y < 60) or (z < 40 and y < 65 and x > 64)): b[83] = 84
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 35 and y < 60) or (z < 40 and y < 65 and x != 40)): b[84] = 85
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 35 and y < 60) or (z < 40 and y != 65 and x > 40)): b[85] = 86
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 35 and y < 60) or (z < 40 and y < 47 and x > 40)): b[86] = 87
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 35 and y < 60) or (z != 40 and y < 65 and x > 40)): b[87] = 88
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 35 and y < 60) or (z < 48 and y < 65 and x > 40)): b[88] = 89
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 35 and y != 60) or (z < 40 and y < 65 and x > 40)): b[89] = 90
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 25 and y < 60) or (z < 40 and y < 65 and x > 40)): b[90] = 91
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 35 and y < 56) or (z < 40 and y < 65 and x > 40)): b[91] = 92
    if (z < 35 and y < 60) or (z < 40 and y < 65 and x > 40):
        pattern_type = 293
    if (y < 40 and x_score > 40) != (y < 40 or x_score > 40): b[92] = 93
    if (y < 40 and x_score > 40) != (y < 49 and x_score > 40): b[93] = 94
    if (y < 40 and x_score > 40) != (y != 40 and x_score > 40): b[94] = 95
    if (y < 40 and x_score > 40) != (y < 40 and x_score > 32): b[95] = 96
    if y < 40 and x_score > 40:
        pattern_type = 294
    if ((x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)) != (
            (x_score != 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)): b[96] = 97
    if ((x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)) != (
            (x_score < 45 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)): b[97] = 98
    if ((x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)) != (
            (x_score < 55 or z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)): b[98] = 99
    if ((x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)) != (
            (x_score < 55 and z < 72 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)): b[99] = 100
    if ((x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)) != (
            (x_score < 55 and z < 65 or (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)): b[100] = 101
    if ((x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)) != (
            (x_score < 55 and z < 65 and (x_score * z) / 84 < 32) or (x_score < 50 and z < 60)): b[101] = 102
    if ((x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)) != (
            (x_score < 55 and z < 65 and (x_score * z) / 100 < 21) or (x_score < 50 and z < 60)): b[102] = 103
    if ((x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)) != (
            (x_score < 55 and z < 65 and (x_score * z) / 100 < 32) and (x_score < 50 and z < 60)): b[103] = 104
    if ((x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)) != (
            (x_score < 55 and z < 65 and (x_score * 63) / 100 < 32) or (x_score < 50 and z < 60)): b[104] = 105
    if ((x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)) != (
            (x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 36 and z != 60)): b[105] = 106
    if ((x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)) != (
            (x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 91)): b[106] = 107
    if ((x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)) != (
            (x_score < 55 and z < 65 and (x_score * z) / 100 < 52) or (x_score < 50 and z < 60)): b[107] = 108
    if (x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60):
        pattern_type = 295
    if (((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 70)) != (
            ((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y != 70)): b[108] = 109
    if (((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 70)) != (
            ((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 77)): b[109] = 110
    if (((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 70)) != (
            ((z ** 2) / (x_score + y + 1) < 28) or (z < 70 or y < 70)): b[110] = 111
    if (((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 70)) != (
            ((z ** 2) / (x_score + y + 1) < 28) or (z != 70 and y < 70)): b[111] = 112
    if (((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 70)) != (
            ((z ** 2) / (x_score + y + 1) < 28) or (z < 56 and y < 70)): b[112] = 113
    if (((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 70)) != (
            ((z ** 2) / (x_score + y + 1) < 28) and (z < 70 and y < 70)): b[113] = 114
    if (((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 70)) != (
            ((z ** 2) / (x_score + y + 1) < 19) or (z < 70 and y < 70)): b[114] = 115
    if (((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 70)) != (
            ((z ** 2) / (x_score + y + 1) != 28) or (z < 70 and y < 70)): b[115] = 116
    if (((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 70)) != (
            ((z ** 2) / (x_score + y + 3) < 28) or (z < 70 and y < 70)): b[116] = 117
    if (((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 70)) != (
            ((z ** 2) / (x_score + 13 + 1) < 28) or (z < 70 and y < 70)): b[117] = 118
    if (((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 70)) != (
            ((z ** 2) / (x_score + y + 1) < 28) or (z < 83 and y < 70)): b[118] = 119
    if ((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 70):
        pattern_type = 296
    if (50 <= x_score < 75 and y > 65) != (50 <= x_score < 75 or y > 65): b[119] = 120
    if (50 <= x_score < 75 and y > 65) != (50 <= x_score < 75 and y != 65): b[120] = 121
    if (50 <= x_score < 75 and y > 65) != (50 <= x_score < 75 and y > 52): b[121] = 122
    if (50 <= x_score < 75 and y > 65) != (50 <= x_score < 83 and y > 65): b[122] = 123
    if (50 <= x_score < 75 and y > 65) != (39 <= x_score < 75 and y > 65): b[123] = 124
    if 50 <= x_score < 75 and y > 65:
        pattern_type = 297
    if ((55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)) != (
            (55 <= y < 80 or z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)): b[124] = 125
    if ((55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)) != (
            (55 <= y < 71 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)): b[125] = 126
    if ((55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)) != (
            (61 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)): b[126] = 127
    if ((55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)) != (
            (55 <= y < 80 and z > 70 or (y * z) / 100 < 52) or (y < 75 and z > 75)): b[127] = 128
    if ((55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)) != (
            (55 <= y < 80 and z > 70 and (y * z) / 100 < 52) and (y < 75 and z > 75)): b[128] = 129
    if ((55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)) != (
            (55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y != 75 and z > 75)): b[129] = 130
    if ((55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)) != (
            (55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 or z > 75)): b[130] = 131
    if ((55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)) != (
            (55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z != 75)): b[131] = 132
    if ((55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)) != (
            (55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 55)): b[132] = 133
    if ((55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)) != (
            (55 <= y < 80 and z > 70 and (y * z) / 89 < 52) or (y < 75 and z > 75)): b[133] = 134
    if ((55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)) != (
            (55 <= y < 80 and z != 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)): b[134] = 135
    if (55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75):
        pattern_type = 298
    if ((60 <= z < 80 and x_score > 70) or (65 <= z < 85 and x_score > 65)) != (
            (60 <= z < 80 or x_score > 70) or (65 <= z < 85 and x_score > 65)): b[135] = 136
    if ((60 <= z < 80 and x_score > 70) or (65 <= z < 85 and x_score > 65)) != (
            (60 <= z < 80 and x_score > 70) and (65 <= z < 85 and x_score > 65)): b[136] = 137
    if ((60 <= z < 80 and x_score > 70) or (65 <= z < 85 and x_score > 65)) != (
            (60 <= z < 80 and x_score > 70) or (65 <= z < 85 or x_score > 65)): b[137] = 138
    if ((60 <= z < 80 and x_score > 70) or (65 <= z < 85 and x_score > 65)) != (
            (60 <= z < 80 and x_score > 70) or (65 <= z < 85 and x_score != 65)): b[138] = 139
    if ((60 <= z < 80 and x_score > 70) or (65 <= z < 85 and x_score > 65)) != (
            (60 <= z < 80 and x_score > 70) or (65 <= z < 85 and x_score > 51)): b[139] = 140
    if ((60 <= z < 80 and x_score > 70) or (65 <= z < 85 and x_score > 65)) != (
            (60 <= z < 80 and x_score != 70) or (65 <= z < 85 and x_score > 65)): b[140] = 141
    if ((60 <= z < 80 and x_score > 70) or (65 <= z < 85 and x_score > 65)) != (
            (60 <= z < 80 and x_score > 61) or (65 <= z < 85 and x_score > 65)): b[141] = 142
    if ((60 <= z < 80 and x_score > 70) or (65 <= z < 85 and x_score > 65)) != (
            (60 <= z < 80 and x_score * 2 > 70) or (65 <= z < 85 and x_score > 65)): b[142] = 143
    if (60 <= z < 80 and x_score > 70) or (65 <= z < 85 and x_score > 65):
        pattern_type = 299
    if (x_score >= 85 and y >= 88 and z >= 85) != (x_score >= 85 or y >= 88 and z >= 85): b[143] = 144
    if (x_score >= 85 and y >= 88 and z >= 85) != (x_score >= 85 and y >= 88 or z >= 85): b[144] = 145
    if (x_score >= 85 and y >= 88 and z >= 85) != (x_score >= 85 and y >= 88 and z != 85): b[145] = 146
    if (x_score >= 85 and y >= 88 and z >= 85) != (x_score >= 85 and y != 88 and z >= 85): b[146] = 147
    if (x_score >= 85 and y >= 88 and z >= 85) != (x_score != 85 and y >= 88 and z >= 85): b[147] = 148
    if (x_score >= 85 and y >= 88 and z >= 85) != (x_score >= 65 and y >= 88 and z >= 85): b[148] = 149
    if (x_score >= 85 and y >= 88 and z >= 85) != (x_score >= 85 and y >= 68 and z >= 85): b[149] = 150
    if (x_score >= 85 and y >= 88 and z >= 85) != (x_score >= 85 and y >= 88 and z >= 67): b[150] = 151
    if (x_score >= 85 and y >= 88 and z >= 85) != (x_score >= 85 and y >= 95 and z >= 85): b[151] = 152
    if x_score >= 85 and y >= 88 and z >= 85:
        pattern_type = 300

    # 返回被触发的规则编号集合
    return set(b.values())


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0


targetPaths = [
    {35, 41, 46, 47, 49, 51, 53, 56, 59, 60, 62, 63, 64, 68, 69, 73, 75, 77, 80, 82, 83, 88, 93, 99, 100, 101, 106, 107,
     111, 112, 116, 119, 125, 128, 131, 132, 133, 136, 138, 139, 141},
    {2, 3, 5, 6, 9, 10, 12, 14, 18, 20, 22, 25, 26, 27, 28, 29, 37, 39, 75, 77, 83, 99, 100, 101, 106, 107, 109, 110,
     111, 116, 120, 125, 128, 131, 132, 133, 135, 136, 138, 139, 141},
    {1, 3, 9, 10, 12, 13, 37, 41, 42, 43, 44, 45, 48, 50, 52, 54, 57, 58, 75, 77, 80, 82, 83, 88, 93, 99, 100, 101, 106,
     107, 111, 116, 119, 128, 131, 132, 133, 136, 138, 139, 141},
    {37, 42, 43, 44, 45, 48, 51, 53, 55, 56, 59, 60, 63, 64, 68, 69, 73, 75, 77, 80, 82, 83, 88, 93, 99, 101, 106, 107,
     111, 112, 116, 119, 125, 128, 131, 133, 136, 138, 139, 141},
    {1, 3, 9, 10, 12, 13, 37, 41, 42, 43, 44, 45, 48, 50, 52, 54, 55, 57, 58, 75, 77, 80, 82, 83, 88, 93, 99, 100, 101,
     106, 107, 113, 114, 128, 131, 132, 133, 136, 138, 139, 141},
    {2, 7, 8, 11, 14, 16, 17, 20, 21, 22, 30, 35, 46, 47, 49, 75, 77, 80, 82, 83, 88, 99, 100, 101, 106, 107, 111, 112,
     116, 119, 125, 128, 131, 132, 133, 136, 138, 139, 141},
    {1, 4, 7, 8, 11, 13, 15, 27, 30, 31, 32, 34, 35, 36, 38, 40, 49, 75, 77, 82, 83, 88, 99, 100, 101, 106, 107, 113,
     114, 125, 128, 131, 132, 133, 135, 136, 138, 139, 141},
    {1, 4, 7, 8, 10, 11, 12, 13, 15, 24, 30, 31, 35, 36, 38, 40, 49, 75, 77, 83, 99, 100, 101, 106, 107, 113, 114, 120,
     125, 128, 131, 132, 133, 135, 136, 138, 139, 141},
    {37, 42, 43, 44, 45, 48, 51, 53, 55, 56, 59, 60, 62, 63, 64, 68, 69, 73, 75, 77, 80, 82, 83, 88, 93, 99, 101, 106,
     107, 111, 112, 116, 119, 129, 136, 138, 139, 141},
    {35, 46, 47, 49, 51, 53, 56, 59, 60, 63, 64, 68, 69, 73, 74, 75, 77, 80, 82, 83, 88, 93, 99, 100, 101, 106, 107,
     113, 114, 128, 131, 132, 133, 136, 138, 139, 141},
    {1, 3, 4, 9, 10, 12, 13, 15, 23, 24, 35, 46, 47, 49, 75, 77, 80, 82, 83, 88, 93, 99, 100, 101, 106, 107, 111, 116,
     119, 128, 131, 132, 133, 136, 138, 139, 141},
    {4, 5, 6, 9, 10, 12, 14, 18, 20, 25, 26, 27, 28, 37, 39, 54, 58, 75, 78, 80, 82, 83, 85, 90, 93, 95, 97, 99, 101,
     120, 121, 122, 125, 128, 131, 132, 135, 138},
    {2, 3, 5, 6, 9, 10, 12, 14, 18, 19, 20, 22, 25, 26, 28, 29, 37, 39, 75, 77, 83, 99, 100, 101, 106, 107, 109, 111,
     116, 120, 125, 128, 135, 136, 138, 139, 141},
    {35, 41, 46, 47, 49, 58, 61, 66, 70, 71, 72, 75, 77, 80, 82, 83, 88, 93, 99, 101, 106, 107, 111, 112, 116, 119, 125,
     128, 131, 132, 133, 136, 138, 139, 141},
    {35, 47, 49, 58, 61, 65, 66, 67, 70, 71, 72, 75, 77, 80, 82, 83, 88, 93, 99, 101, 106, 107, 111, 112, 116, 119, 125,
     128, 131, 132, 133, 136, 138, 139, 141},
    {2, 3, 5, 6, 9, 14, 17, 18, 20, 21, 22, 27, 30, 35, 36, 38, 40, 49, 75, 77, 83, 99, 101, 106, 107, 111, 112, 116,
     119, 120, 129, 134, 136, 138, 139, 141},
    {1, 4, 7, 8, 11, 13, 15, 25, 26, 28, 29, 37, 39, 99, 100, 101, 107, 109, 110, 111, 116, 120, 124, 125, 128, 131,
     132, 133, 135, 136, 138, 139, 141, 143},
    {1, 4, 7, 8, 11, 13, 15, 25, 26, 28, 29, 37, 39, 93, 95, 98, 102, 103, 104, 109, 110, 111, 116, 120, 124, 125, 128,
     131, 132, 133, 135, 136, 141, 143},
    {27, 31, 32, 35, 36, 40, 41, 42, 46, 47, 49, 50, 53, 55, 60, 61, 62, 64, 68, 69, 72, 73, 80, 93, 94, 95, 111, 112,
     116, 129, 136, 138, 144, 145, 147},
    {1, 2, 7, 8, 11, 13, 15, 24, 25, 26, 27, 28, 37, 39, 54, 58, 75, 78, 81, 92, 93, 95, 97, 99, 101, 120, 121, 122,
     125, 128, 131, 132, 135, 136, 138},
    {1, 4, 8, 10, 11, 12, 13, 15, 27, 30, 35, 38, 49, 75, 77, 80, 82, 83, 88, 99, 100, 101, 106, 107, 111, 112, 116,
     119, 127, 129, 136, 138, 139, 141},
    {1, 2, 7, 8, 11, 13, 15, 23, 24, 26, 27, 28, 37, 39, 54, 58, 75, 78, 80, 82, 90, 93, 95, 97, 99, 101, 120, 123, 125,
     128, 131, 132, 135, 136, 138},
    {27, 30, 31, 32, 35, 36, 40, 49, 58, 65, 71, 80, 93, 94, 95, 111, 112, 116, 119, 120, 121, 125, 128, 131, 133, 136,
     138, 139, 140, 141, 142, 143},
    {2, 4, 5, 6, 8, 9, 10, 11, 12, 14, 20, 22, 25, 26, 27, 28, 32, 37, 39, 40, 54, 58, 75, 78, 81, 93, 94, 95, 98, 104,
     105, 120, 121, 128, 131, 132},
    {2, 3, 5, 6, 9, 14, 17, 18, 19, 20, 21, 22, 27, 30, 31, 32, 35, 36, 38, 40, 49, 75, 77, 83, 99, 101, 106, 107, 116,
     120, 125, 128, 130, 131, 145},
    {27, 35, 49, 58, 61, 65, 66, 67, 70, 71, 72, 80, 93, 96, 99, 101, 106, 107, 111, 112, 116, 119, 125, 128, 131, 132,
     133, 136, 138, 139, 141},
    {1, 2, 7, 8, 13, 27, 28, 37, 39, 51, 52, 54, 56, 58, 59, 75, 78, 81, 91, 92, 93, 95, 97, 99, 101, 125, 128, 131,
     132, 135, 136, 138, 144},
    {3, 4, 5, 6, 9, 10, 12, 14, 19, 27, 28, 37, 39, 54, 58, 75, 78, 80, 82, 90, 93, 95, 97, 99, 101, 114, 120, 128, 136,
     138, 144, 145, 146},
    {1, 2, 3, 13, 25, 26, 27, 28, 31, 32, 33, 37, 38, 40, 52, 54, 55, 56, 57, 58, 59, 75, 78, 81, 97, 99, 101, 128, 131,
     132, 136, 138, 144},
    {2, 3, 5, 6, 9, 10, 12, 14, 18, 19, 20, 22, 25, 26, 28, 29, 37, 39, 75, 77, 83, 99, 101, 106, 107, 116, 120, 130,
     131, 144, 145, 148},
    {2, 3, 5, 6, 9, 10, 12, 14, 18, 19, 20, 22, 25, 26, 28, 37, 39, 93, 95, 98, 102, 103, 104, 114, 115, 118, 120, 124,
     136, 141, 143},
    {1, 4, 7, 8, 11, 13, 15, 25, 26, 28, 29, 37, 39, 75, 77, 83, 99, 100, 101, 106, 107, 116, 120, 126, 129, 134, 136,
     138, 139, 141},
    {27, 31, 32, 35, 36, 40, 49, 58, 63, 65, 66, 70, 71, 80, 93, 94, 95, 111, 112, 116, 119, 120, 121, 125, 128, 131,
     132, 133, 137},
    {1, 4, 7, 8, 11, 13, 15, 25, 26, 28, 29, 37, 39, 93, 95, 99, 101, 108, 113, 114, 125, 128, 131, 132, 133, 135, 136,
     141, 143},
    {2, 3, 5, 6, 9, 10, 12, 14, 18, 20, 22, 25, 26, 27, 28, 37, 39, 83, 109, 110, 111, 116, 117, 120, 125, 128, 131,
     132, 135},
    {2, 5, 6, 9, 10, 12, 14, 18, 20, 22, 30, 35, 36, 38, 40, 49, 76, 79, 80, 82, 83, 88, 114, 125, 128, 131, 132, 135},
    {1, 3, 13, 35, 49, 50, 51, 53, 56, 60, 63, 64, 68, 69, 73, 74, 75, 78, 80, 82, 83, 88, 89, 93, 114, 128, 131, 132},
    {50, 53, 55, 56, 59, 60, 61, 62, 68, 69, 72, 93, 95, 116, 120, 125, 128, 130, 131, 136, 138, 144, 145, 147, 150},
    {2, 6, 9, 10, 12, 14, 20, 22, 25, 26, 27, 28, 31, 32, 33, 37, 39, 40, 79, 81, 84, 87, 125, 128, 131, 132, 135},
    {8, 10, 12, 25, 26, 27, 28, 37, 39, 54, 58, 79, 80, 82, 83, 86, 90, 114, 120, 125, 128, 131, 132, 135},
    {1, 7, 8, 11, 13, 51, 52, 54, 58, 93, 95, 116, 120, 130, 131, 137, 144, 145, 146, 151},
    {1, 7, 8, 11, 13, 93, 95, 116, 120, 130, 131, 136, 138, 144, 145, 148, 149},
    {1, 7, 8, 11, 13, 93, 95, 116, 120, 130, 131, 136, 138, 152},
]


class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 2000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(targetPaths)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }


def compute_robustness(state, path):
    base = execute_Tr(*state)
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                        np.clip(state[0] + dx, STATE_MIN_X, STATE_MAX_X),
                        np.clip(state[1] + dy, STATE_MIN_Y, STATE_MAX_Y),
                        np.clip(state[2] + dz, STATE_MIN_Z, STATE_MAX_Z)
                    ])
                n_trig = execute_Tr(*neighbor)
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = targetPaths[target_path_idx]
    samples = []
    attempts = 0

    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(STATE_MIN_X, STATE_MAX_X),
            random.randint(STATE_MIN_Y, STATE_MAX_Y),
            random.randint(STATE_MIN_Z, STATE_MAX_Z)
        ])
        triggered = execute_Tr(*state)

        if not triggered:
            continue

        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)

        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }

        samples.append(sample_data)

    return samples


def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample[
                'robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)

        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]

    return selected_samples


def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = targetPaths[target_path_idx]
        samples = []
        attempts = 0

        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(STATE_MIN_X, STATE_MAX_X),
                random.randint(STATE_MIN_Y, STATE_MAX_Y),
                random.randint(STATE_MIN_Z, STATE_MAX_Z)
            ])
            triggered = execute_Tr(*state)

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)

            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)

        random.shuffle(samples)
        return samples[:config.top_k_samples]

    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")

        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)


def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}

    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates

    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}

        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config,
                                                         shared_candidates[path_idx])

            strategy_results[path_idx] = samples

        results[strategy_name] = strategy_results

    return results


def analyze_fitness_values(results, config):
    analysis_results = {}

    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []

        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])

        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }

        all_scores = []

        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])

        scores_array = np.array(all_scores)

        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)

        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })

        analysis_results[strategy_name] = analysis

    return analysis_results


def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []

    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)

    df = pd.DataFrame(df_data)
    return df, analysis_results


def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []

    print(f"Starting {num_runs} experiments...")

    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")

        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)

        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)

        print(f"Experiment {run_idx + 1} completed.")

    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df


def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')

    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width

    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []

    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]

        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)

    stats_df = pd.DataFrame(stats_data)

    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Results saved to: {output_path}")


def main():
    results_df = run_multiple_experiments(num_runs=20)

    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")

    save_results_to_excel(results_df, output_path)

    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)

    return results_df, output_path


if __name__ == "__main__":
    results, output_file = main()
