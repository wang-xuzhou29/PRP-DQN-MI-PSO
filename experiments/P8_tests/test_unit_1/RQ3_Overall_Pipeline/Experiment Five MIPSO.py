import os
import random
import numpy as np
import time
import psutil
import math
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

NUM_RUNS = 20  # 默认实验运行次数

# === 新的三维范围设置 ===
LIGHT_MIN = 2
LIGHT_MAX = 100
MOISTURE_MIN = 2
MOISTURE_MAX = 100
TEMP_MIN = 2
TEMP_MAX = 100

BOUNDS = {
    'light': (LIGHT_MIN, LIGHT_MAX),
    'temp': (TEMP_MIN, TEMP_MAX),
    'moisture': (MOISTURE_MIN, MOISTURE_MAX)
}


def clip_state(state):
    """将状态限制在各维度的边界内"""
    return np.array([
        np.clip(state[0], BOUNDS['light'][0], BOUNDS['light'][1]),
        np.clip(state[1], BOUNDS['temp'][0], BOUNDS['temp'][1]),
        np.clip(state[2], BOUNDS['moisture'][0], BOUNDS['moisture'][1])
    ])


def execute_Tr(position):
    """执行目标函数并返回触发的路径"""
    x = int(np.clip(position[0], BOUNDS['light'][0], BOUNDS['light'][1]))
    temp = int(np.clip(position[1], BOUNDS['temp'][0], BOUNDS['temp'][1]))
    z = int(np.clip(position[2], BOUNDS['moisture'][0], BOUNDS['moisture'][1]))
    return category1_multivariable_control(x, temp, z)


# === 目标函数 ===
def category1_multivariable_control(a):
    """
    替换原有的编排规则函数
    参数a: 包含3个元素的元组或数组，分别对应path_depth, file_count, access_level
    返回: 触发的规则编号集合
    """
    path_depth, file_count, access_level = float(a[0]), int(a[1]), float(a[2])

    # 使用正确的变量名
    x, y, z = path_depth, file_count, access_level

    triggered = set()
    # 创建一个字典来存储b数组的值，用于跟踪哪些规则被触发
    b = {}
    x_score = max(0, 100 - x * 1.67)

    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2): b[
        0] = 1
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((27 * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[1] = 2
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * 63) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[2] = 3
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (36 + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[3] = 4
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 7) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[4] = 5
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 12) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[5] = 6
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[6] = 7
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((z * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[7] = 8
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * x) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[8] = 9
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[9] = 10
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (x + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[10] = 11
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (y + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[11] = 12
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) != 50 and x ** 2 + y ** 2 > z ** 2):
        b[12] = 13
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 70 and x ** 2 + y ** 2 > z ** 2):
        b[13] = 14
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 28 and x ** 2 + y ** 2 > z ** 2):
        b[14] = 15
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and y ** 2 + y ** 2 > z ** 2):
        b[15] = 16
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 1 + y ** 2 > z ** 2):
        b[16] = 17
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 1 > z ** 2):
        b[17] = 18
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + x ** 2 > z ** 2):
        b[18] = 19
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 3):
        b[19] = 20
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x * 2 + y ** 2 > z ** 2): b[
        20] = 21
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2.5):
        b[21] = 22
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z * 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[22] = 23
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z - 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[23] = 24
    if (x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2:
        pattern_type = "291A1"
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 3 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[24] = 25
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 4 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[25] = 26
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - z ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[26] = 27
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 1) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[27] = 28
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[28] = 29
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 1 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[29] = 30
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (x + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[30] = 31
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[31] = 32
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 1.5) < -30 or (abs(x - y) * z) / 100 > 45): b[32] = 33
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z - 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[33] = 34
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) != -30 or (abs(x - y) * z) / 100 > 45): b[34] = 35
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -10 or (abs(x - y) * z) / 100 > 45): b[35] = 36
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 and (abs(x - y) * z) / 100 > 45): b[36] = 37
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (24 ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[37] = 38
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - 19 ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[38] = 39
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (42 + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[39] = 40
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45): b[40] = 41
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - z) * z) / 100 > 45): b[41] = 42
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(52 - y) * z) / 100 > 45): b[42] = 43
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - 42) * z) / 100 > 45): b[43] = 44
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * 55) / 100 > 45): b[44] = 45
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 75 > 45): b[45] = 46
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 25): b[46] = 47
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 65): b[47] = 48
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 != 45): b[48] = 49
    if (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45:
        pattern_type = "291A2"
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((46 + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35): b[49] = 50
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + 37) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35): b[50] = 51
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 3) ** 2 < z * 20 and (x * y * z) / 1000 < 35): b[51] = 52
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 3 < z * 20 and (x * y * z) / 1000 < 35): b[52] = 53
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 != z * 20 and (x * y * z) / 1000 < 35): b[53] = 54
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < x * 20 and (x * y * z) / 1000 < 35): b[54] = 55
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35): b[55] = 56
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 30 and (x * y * z) / 1000 < 35): b[56] = 57
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 or (x * y * z) / 1000 < 35): b[57] = 58
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < 52 * 20 and (x * y * z) / 1000 < 35): b[58] = 59
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 3 < z * 20 and (x * y * z) / 1000 < 35): b[59] = 60
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (y * y * z) / 1000 < 35): b[60] = 61
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (z * y * z) / 1000 < 35): b[61] = 62
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35): b[62] = 63
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * z * z) / 1000 < 35): b[63] = 64
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35): b[64] = 65
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * y * y) / 1000 < 35): b[65] = 66
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1200 < 35): b[66] = 67
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 700 < 35): b[67] = 68
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 15): b[68] = 69
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 55): b[69] = 70
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 != 35): b[70] = 71
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (34 * y * z) / 1000 < 35): b[71] = 72
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * 53 * z) / 1000 < 35): b[72] = 73
    if (((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * y * 72) / 1000 < 35): b[73] = 74
    if ((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35:
        pattern_type = "291A3"
    if (x_score < 30 and z < 50) != (x_score < 30 or z < 50): b[74] = 75
    if (x_score < 30 and z < 50) != (x_score < 30 and z < 40): b[75] = 76
    if (x_score < 30 and z < 50) != (x_score < 30 and z != 50): b[76] = 77
    if (x_score < 30 and z < 50) != (x_score != 30 and z < 50): b[77] = 78
    if (x_score < 30 and z < 50) != (x_score < 19 and z < 50): b[78] = 79
    if x_score < 30 and z < 50:
        pattern_type = 292
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 35 or y < 60) or (z < 40 and y < 65 and x > 40)): b[79] = 80
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 35 and y < 60) and (z < 40 and y < 65 and x > 40)): b[80] = 81
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 35 and y < 60) or (z < 40 or y < 65 and x > 40)): b[81] = 82
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 35 and y < 60) or (z < 40 and y < 65 or x > 40)): b[82] = 83
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 35 and y < 60) or (z < 40 and y < 65 and x > 64)): b[83] = 84
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 35 and y < 60) or (z < 40 and y < 65 and x != 40)): b[84] = 85
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 35 and y < 60) or (z < 40 and y != 65 and x > 40)): b[85] = 86
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 35 and y < 60) or (z < 40 and y < 47 and x > 40)): b[86] = 87
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 35 and y < 60) or (z != 40 and y < 65 and x > 40)): b[87] = 88
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 35 and y < 60) or (z < 48 and y < 65 and x > 40)): b[88] = 89
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 35 and y != 60) or (z < 40 and y < 65 and x > 40)): b[89] = 90
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 25 and y < 60) or (z < 40 and y < 65 and x > 40)): b[90] = 91
    if ((z < 35 and y < 60) or (z < 40 and y < 65 and x > 40)) != (
            (z < 35 and y < 56) or (z < 40 and y < 65 and x > 40)): b[91] = 92
    if (z < 35 and y < 60) or (z < 40 and y < 65 and x > 40):
        pattern_type = 293
    if (y < 40 and x_score > 40) != (y < 40 or x_score > 40): b[92] = 93
    if (y < 40 and x_score > 40) != (y < 49 and x_score > 40): b[93] = 94
    if (y < 40 and x_score > 40) != (y != 40 and x_score > 40): b[94] = 95
    if (y < 40 and x_score > 40) != (y < 40 and x_score > 32): b[95] = 96
    if y < 40 and x_score > 40:
        pattern_type = 294
    if ((x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)) != (
            (x_score != 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)): b[96] = 97
    if ((x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)) != (
            (x_score < 45 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)): b[97] = 98
    if ((x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)) != (
            (x_score < 55 or z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)): b[98] = 99
    if ((x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)) != (
            (x_score < 55 and z < 72 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)): b[99] = 100
    if ((x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)) != (
            (x_score < 55 and z < 65 or (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)): b[100] = 101
    if ((x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)) != (
            (x_score < 55 and z < 65 and (x_score * z) / 84 < 32) or (x_score < 50 and z < 60)): b[101] = 102
    if ((x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)) != (
            (x_score < 55 and z < 65 and (x_score * z) / 100 < 21) or (x_score < 50 and z < 60)): b[102] = 103
    if ((x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)) != (
            (x_score < 55 and z < 65 and (x_score * z) / 100 < 32) and (x_score < 50 and z < 60)): b[103] = 104
    if ((x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)) != (
            (x_score < 55 and z < 65 and (x_score * 63) / 100 < 32) or (x_score < 50 and z < 60)): b[104] = 105
    if ((x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)) != (
            (x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 36 and z != 60)): b[105] = 106
    if ((x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)) != (
            (x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 91)): b[106] = 107
    if ((x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60)) != (
            (x_score < 55 and z < 65 and (x_score * z) / 100 < 52) or (x_score < 50 and z < 60)): b[107] = 108
    if (x_score < 55 and z < 65 and (x_score * z) / 100 < 32) or (x_score < 50 and z < 60):
        pattern_type = 295
    if (((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 70)) != (
            ((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y != 70)): b[108] = 109
    if (((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 70)) != (
            ((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 77)): b[109] = 110
    if (((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 70)) != (
            ((z ** 2) / (x_score + y + 1) < 28) or (z < 70 or y < 70)): b[110] = 111
    if (((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 70)) != (
            ((z ** 2) / (x_score + y + 1) < 28) or (z != 70 and y < 70)): b[111] = 112
    if (((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 70)) != (
            ((z ** 2) / (x_score + y + 1) < 28) or (z < 56 and y < 70)): b[112] = 113
    if (((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 70)) != (
            ((z ** 2) / (x_score + y + 1) < 28) and (z < 70 and y < 70)): b[113] = 114
    if (((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 70)) != (
            ((z ** 2) / (x_score + y + 1) < 19) or (z < 70 and y < 70)): b[114] = 115
    if (((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 70)) != (
            ((z ** 2) / (x_score + y + 1) != 28) or (z < 70 and y < 70)): b[115] = 116
    if (((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 70)) != (
            ((z ** 2) / (x_score + y + 3) < 28) or (z < 70 and y < 70)): b[116] = 117
    if (((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 70)) != (
            ((z ** 2) / (x_score + 13 + 1) < 28) or (z < 70 and y < 70)): b[117] = 118
    if (((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 70)) != (
            ((z ** 2) / (x_score + y + 1) < 28) or (z < 83 and y < 70)): b[118] = 119
    if ((z ** 2) / (x_score + y + 1) < 28) or (z < 70 and y < 70):
        pattern_type = 296
    if (50 <= x_score < 75 and y > 65) != (50 <= x_score < 75 or y > 65): b[119] = 120
    if (50 <= x_score < 75 and y > 65) != (50 <= x_score < 75 and y != 65): b[120] = 121
    if (50 <= x_score < 75 and y > 65) != (50 <= x_score < 75 and y > 52): b[121] = 122
    if (50 <= x_score < 75 and y > 65) != (50 <= x_score < 83 and y > 65): b[122] = 123
    if (50 <= x_score < 75 and y > 65) != (39 <= x_score < 75 and y > 65): b[123] = 124
    if 50 <= x_score < 75 and y > 65:
        pattern_type = 297
    if ((55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)) != (
            (55 <= y < 80 or z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)): b[124] = 125
    if ((55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)) != (
            (55 <= y < 71 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)): b[125] = 126
    if ((55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)) != (
            (61 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)): b[126] = 127
    if ((55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)) != (
            (55 <= y < 80 and z > 70 or (y * z) / 100 < 52) or (y < 75 and z > 75)): b[127] = 128
    if ((55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)) != (
            (55 <= y < 80 and z > 70 and (y * z) / 100 < 52) and (y < 75 and z > 75)): b[128] = 129
    if ((55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)) != (
            (55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y != 75 and z > 75)): b[129] = 130
    if ((55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)) != (
            (55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 or z > 75)): b[130] = 131
    if ((55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)) != (
            (55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z != 75)): b[131] = 132
    if ((55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)) != (
            (55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 55)): b[132] = 133
    if ((55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)) != (
            (55 <= y < 80 and z > 70 and (y * z) / 89 < 52) or (y < 75 and z > 75)): b[133] = 134
    if ((55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)) != (
            (55 <= y < 80 and z != 70 and (y * z) / 100 < 52) or (y < 75 and z > 75)): b[134] = 135
    if (55 <= y < 80 and z > 70 and (y * z) / 100 < 52) or (y < 75 and z > 75):
        pattern_type = 298
    if ((60 <= z < 80 and x_score > 70) or (65 <= z < 85 and x_score > 65)) != (
            (60 <= z < 80 or x_score > 70) or (65 <= z < 85 and x_score > 65)): b[135] = 136
    if ((60 <= z < 80 and x_score > 70) or (65 <= z < 85 and x_score > 65)) != (
            (60 <= z < 80 and x_score > 70) and (65 <= z < 85 and x_score > 65)): b[136] = 137
    if ((60 <= z < 80 and x_score > 70) or (65 <= z < 85 and x_score > 65)) != (
            (60 <= z < 80 and x_score > 70) or (65 <= z < 85 or x_score > 65)): b[137] = 138
    if ((60 <= z < 80 and x_score > 70) or (65 <= z < 85 and x_score > 65)) != (
            (60 <= z < 80 and x_score > 70) or (65 <= z < 85 and x_score != 65)): b[138] = 139
    if ((60 <= z < 80 and x_score > 70) or (65 <= z < 85 and x_score > 65)) != (
            (60 <= z < 80 and x_score > 70) or (65 <= z < 85 and x_score > 51)): b[139] = 140
    if ((60 <= z < 80 and x_score > 70) or (65 <= z < 85 and x_score > 65)) != (
            (60 <= z < 80 and x_score != 70) or (65 <= z < 85 and x_score > 65)): b[140] = 141
    if ((60 <= z < 80 and x_score > 70) or (65 <= z < 85 and x_score > 65)) != (
            (60 <= z < 80 and x_score > 61) or (65 <= z < 85 and x_score > 65)): b[141] = 142
    if ((60 <= z < 80 and x_score > 70) or (65 <= z < 85 and x_score > 65)) != (
            (60 <= z < 80 and x_score * 2 > 70) or (65 <= z < 85 and x_score > 65)): b[142] = 143
    if (60 <= z < 80 and x_score > 70) or (65 <= z < 85 and x_score > 65):
        pattern_type = 299
    if (x_score >= 85 and y >= 88 and z >= 85) != (x_score >= 85 or y >= 88 and z >= 85): b[143] = 144
    if (x_score >= 85 and y >= 88 and z >= 85) != (x_score >= 85 and y >= 88 or z >= 85): b[144] = 145
    if (x_score >= 85 and y >= 88 and z >= 85) != (x_score >= 85 and y >= 88 and z != 85): b[145] = 146
    if (x_score >= 85 and y >= 88 and z >= 85) != (x_score >= 85 and y != 88 and z >= 85): b[146] = 147
    if (x_score >= 85 and y >= 88 and z >= 85) != (x_score != 85 and y >= 88 and z >= 85): b[147] = 148
    if (x_score >= 85 and y >= 88 and z >= 85) != (x_score >= 65 and y >= 88 and z >= 85): b[148] = 149
    if (x_score >= 85 and y >= 88 and z >= 85) != (x_score >= 85 and y >= 68 and z >= 85): b[149] = 150
    if (x_score >= 85 and y >= 88 and z >= 85) != (x_score >= 85 and y >= 88 and z >= 67): b[150] = 151
    if (x_score >= 85 and y >= 88 and z >= 85) != (x_score >= 85 and y >= 95 and z >= 85): b[151] = 152
    if x_score >= 85 and y >= 88 and z >= 85:
        pattern_type = 300

    # 返回被触发的规则编号集合
    return set(b.values())


# 目标路径定义
targetPaths = [
    {35, 41, 46, 47, 49, 51, 53, 56, 59, 60, 62, 63, 64, 68, 69, 73, 75, 77, 80, 82, 83, 88, 93, 99, 100, 101, 106, 107,
     111, 112, 116, 119, 125, 128, 131, 132, 133, 136, 138, 139, 141},
    {2, 3, 5, 6, 9, 10, 12, 14, 18, 20, 22, 25, 26, 27, 28, 29, 37, 39, 75, 77, 83, 99, 100, 101, 106, 107, 109, 110,
     111, 116, 120, 125, 128, 131, 132, 133, 135, 136, 138, 139, 141},
    {1, 3, 9, 10, 12, 13, 37, 41, 42, 43, 44, 45, 48, 50, 52, 54, 57, 58, 75, 77, 80, 82, 83, 88, 93, 99, 100, 101, 106,
     107, 111, 116, 119, 128, 131, 132, 133, 136, 138, 139, 141},
    {37, 42, 43, 44, 45, 48, 51, 53, 55, 56, 59, 60, 63, 64, 68, 69, 73, 75, 77, 80, 82, 83, 88, 93, 99, 101, 106, 107,
     111, 112, 116, 119, 125, 128, 131, 133, 136, 138, 139, 141},
    {1, 3, 9, 10, 12, 13, 37, 41, 42, 43, 44, 45, 48, 50, 52, 54, 55, 57, 58, 75, 77, 80, 82, 83, 88, 93, 99, 100, 101,
     106, 107, 113, 114, 128, 131, 132, 133, 136, 138, 139, 141},
    {2, 7, 8, 11, 14, 16, 17, 20, 21, 22, 30, 35, 46, 47, 49, 75, 77, 80, 82, 83, 88, 99, 100, 101, 106, 107, 111, 112,
     116, 119, 125, 128, 131, 132, 133, 136, 138, 139, 141},
    {1, 4, 7, 8, 11, 13, 15, 27, 30, 31, 32, 34, 35, 36, 38, 40, 49, 75, 77, 82, 83, 88, 99, 100, 101, 106, 107, 113,
     114, 125, 128, 131, 132, 133, 135, 136, 138, 139, 141},
    {1, 4, 7, 8, 10, 11, 12, 13, 15, 24, 30, 31, 35, 36, 38, 40, 49, 75, 77, 83, 99, 100, 101, 106, 107, 113, 114, 120,
     125, 128, 131, 132, 133, 135, 136, 138, 139, 141},
    {37, 42, 43, 44, 45, 48, 51, 53, 55, 56, 59, 60, 62, 63, 64, 68, 69, 73, 75, 77, 80, 82, 83, 88, 93, 99, 101, 106,
     107, 111, 112, 116, 119, 129, 136, 138, 139, 141},
    {35, 46, 47, 49, 51, 53, 56, 59, 60, 63, 64, 68, 69, 73, 74, 75, 77, 80, 82, 83, 88, 93, 99, 100, 101, 106, 107,
     113, 114, 128, 131, 132, 133, 136, 138, 139, 141},
    {1, 3, 4, 9, 10, 12, 13, 15, 23, 24, 35, 46, 47, 49, 75, 77, 80, 82, 83, 88, 93, 99, 100, 101, 106, 107, 111, 116,
     119, 128, 131, 132, 133, 136, 138, 139, 141},
    {4, 5, 6, 9, 10, 12, 14, 18, 20, 25, 26, 27, 28, 37, 39, 54, 58, 75, 78, 80, 82, 83, 85, 90, 93, 95, 97, 99, 101,
     120, 121, 122, 125, 128, 131, 132, 135, 138},
    {2, 3, 5, 6, 9, 10, 12, 14, 18, 19, 20, 22, 25, 26, 28, 29, 37, 39, 75, 77, 83, 99, 100, 101, 106, 107, 109, 111,
     116, 120, 125, 128, 135, 136, 138, 139, 141},
    {35, 41, 46, 47, 49, 58, 61, 66, 70, 71, 72, 75, 77, 80, 82, 83, 88, 93, 99, 101, 106, 107, 111, 112, 116, 119, 125,
     128, 131, 132, 133, 136, 138, 139, 141},
    {35, 47, 49, 58, 61, 65, 66, 67, 70, 71, 72, 75, 77, 80, 82, 83, 88, 93, 99, 101, 106, 107, 111, 112, 116, 119, 125,
     128, 131, 132, 133, 136, 138, 139, 141},
    {2, 3, 5, 6, 9, 14, 17, 18, 20, 21, 22, 27, 30, 35, 36, 38, 40, 49, 75, 77, 83, 99, 101, 106, 107, 111, 112, 116,
     119, 120, 129, 134, 136, 138, 139, 141},
    {1, 4, 7, 8, 11, 13, 15, 25, 26, 28, 29, 37, 39, 99, 100, 101, 107, 109, 110, 111, 116, 120, 124, 125, 128, 131,
     132, 133, 135, 136, 138, 139, 141, 143},
    {1, 4, 7, 8, 11, 13, 15, 25, 26, 28, 29, 37, 39, 93, 95, 98, 102, 103, 104, 109, 110, 111, 116, 120, 124, 125, 128,
     131, 132, 133, 135, 136, 141, 143},
    {27, 31, 32, 35, 36, 40, 41, 42, 46, 47, 49, 50, 53, 55, 60, 61, 62, 64, 68, 69, 72, 73, 80, 93, 94, 95, 111, 112,
     116, 129, 136, 138, 144, 145, 147},
    {1, 2, 7, 8, 11, 13, 15, 24, 25, 26, 27, 28, 37, 39, 54, 58, 75, 78, 81, 92, 93, 95, 97, 99, 101, 120, 121, 122,
     125, 128, 131, 132, 135, 136, 138},
    {1, 4, 8, 10, 11, 12, 13, 15, 27, 30, 35, 38, 49, 75, 77, 80, 82, 83, 88, 99, 100, 101, 106, 107, 111, 112, 116,
     119, 127, 129, 136, 138, 139, 141},
    {1, 2, 7, 8, 11, 13, 15, 23, 24, 26, 27, 28, 37, 39, 54, 58, 75, 78, 80, 82, 90, 93, 95, 97, 99, 101, 120, 123, 125,
     128, 131, 132, 135, 136, 138},
    {27, 30, 31, 32, 35, 36, 40, 49, 58, 65, 71, 80, 93, 94, 95, 111, 112, 116, 119, 120, 121, 125, 128, 131, 133, 136,
     138, 139, 140, 141, 142, 143},
    {2, 4, 5, 6, 8, 9, 10, 11, 12, 14, 20, 22, 25, 26, 27, 28, 32, 37, 39, 40, 54, 58, 75, 78, 81, 93, 94, 95, 98, 104,
     105, 120, 121, 128, 131, 132},
    {2, 3, 5, 6, 9, 14, 17, 18, 19, 20, 21, 22, 27, 30, 31, 32, 35, 36, 38, 40, 49, 75, 77, 83, 99, 101, 106, 107, 116,
     120, 125, 128, 130, 131, 145},
    {27, 35, 49, 58, 61, 65, 66, 67, 70, 71, 72, 80, 93, 96, 99, 101, 106, 107, 111, 112, 116, 119, 125, 128, 131, 132,
     133, 136, 138, 139, 141},
    {1, 2, 7, 8, 13, 27, 28, 37, 39, 51, 52, 54, 56, 58, 59, 75, 78, 81, 91, 92, 93, 95, 97, 99, 101, 125, 128, 131,
     132, 135, 136, 138, 144},
    {3, 4, 5, 6, 9, 10, 12, 14, 19, 27, 28, 37, 39, 54, 58, 75, 78, 80, 82, 90, 93, 95, 97, 99, 101, 114, 120, 128, 136,
     138, 144, 145, 146},
    {1, 2, 3, 13, 25, 26, 27, 28, 31, 32, 33, 37, 38, 40, 52, 54, 55, 56, 57, 58, 59, 75, 78, 81, 97, 99, 101, 128, 131,
     132, 136, 138, 144},
    {2, 3, 5, 6, 9, 10, 12, 14, 18, 19, 20, 22, 25, 26, 28, 29, 37, 39, 75, 77, 83, 99, 101, 106, 107, 116, 120, 130,
     131, 144, 145, 148},
    {2, 3, 5, 6, 9, 10, 12, 14, 18, 19, 20, 22, 25, 26, 28, 37, 39, 93, 95, 98, 102, 103, 104, 114, 115, 118, 120, 124,
     136, 141, 143},
    {1, 4, 7, 8, 11, 13, 15, 25, 26, 28, 29, 37, 39, 75, 77, 83, 99, 100, 101, 106, 107, 116, 120, 126, 129, 134, 136,
     138, 139, 141},
    {27, 31, 32, 35, 36, 40, 49, 58, 63, 65, 66, 70, 71, 80, 93, 94, 95, 111, 112, 116, 119, 120, 121, 125, 128, 131,
     132, 133, 137},
    {1, 4, 7, 8, 11, 13, 15, 25, 26, 28, 29, 37, 39, 93, 95, 99, 101, 108, 113, 114, 125, 128, 131, 132, 133, 135, 136,
     141, 143},
    {2, 3, 5, 6, 9, 10, 12, 14, 18, 20, 22, 25, 26, 27, 28, 37, 39, 83, 109, 110, 111, 116, 117, 120, 125, 128, 131,
     132, 135},
    {2, 5, 6, 9, 10, 12, 14, 18, 20, 22, 30, 35, 36, 38, 40, 49, 76, 79, 80, 82, 83, 88, 114, 125, 128, 131, 132, 135},
    {1, 3, 13, 35, 49, 50, 51, 53, 56, 60, 63, 64, 68, 69, 73, 74, 75, 78, 80, 82, 83, 88, 89, 93, 114, 128, 131, 132},
    {50, 53, 55, 56, 59, 60, 61, 62, 68, 69, 72, 93, 95, 116, 120, 125, 128, 130, 131, 136, 138, 144, 145, 147, 150},
    {2, 6, 9, 10, 12, 14, 20, 22, 25, 26, 27, 28, 31, 32, 33, 37, 39, 40, 79, 81, 84, 87, 125, 128, 131, 132, 135},
    {8, 10, 12, 25, 26, 27, 28, 37, 39, 54, 58, 79, 80, 82, 83, 86, 90, 114, 120, 125, 128, 131, 132, 135},
    {1, 7, 8, 11, 13, 51, 52, 54, 58, 93, 95, 116, 120, 130, 131, 137, 144, 145, 146, 151},
    {1, 7, 8, 11, 13, 93, 95, 116, 120, 130, 131, 136, 138, 144, 145, 148, 149},
    {1, 7, 8, 11, 13, 93, 95, 116, 120, 130, 131, 136, 138, 152},
]


# === 增强版指标收集器 ===
class MetricsCollector:
    def __init__(self):
        self.pso_start_time = None
        self.pso_end_time = None
        self.perfect_solutions_count = 0
        self.total_paths_count = 0
        self.final_fitness_scores = []
        self.pso_convergence_iterations = []
        self.pso_reset_counts = []
        self.path_execution_times = []

    def start_pso_phase(self):
        self.pso_start_time = time.time()

    def end_pso_phase(self):
        self.pso_end_time = time.time()

    def record_pso_result(self, fitness, is_perfect_match, convergence_iter=None, path_id=None, method='MI-PSO',
                          reset_count=0, execution_time=0):
        self.final_fitness_scores.append(fitness)
        self.total_paths_count += 1
        self.pso_reset_counts.append(reset_count)
        self.path_execution_times.append(execution_time)

        if is_perfect_match:
            self.perfect_solutions_count += 1
        if convergence_iter is not None:
            self.pso_convergence_iterations.append(convergence_iter)


# === 粒子类 ===
class Particle:
    def __init__(self):
        self.position = np.array([
            np.random.uniform(BOUNDS['light'][0], BOUNDS['light'][1]),
            np.random.uniform(BOUNDS['temp'][0], BOUNDS['temp'][1]),
            np.random.uniform(BOUNDS['moisture'][0], BOUNDS['moisture'][1])
        ])
        self.velocity = np.array([
            np.random.uniform(-5, 5),
            np.random.uniform(-3, 3),
            np.random.uniform(-5, 5)
        ])
        self.best_position = self.position.copy()
        self.best_fitness = 0
        self.fitness = 0


# === 单独的 MI-PSO 优化器类（已完全修复随机性重评估Bug） ===
class MIPSO:
    def __init__(self, target_path, swarm_size=20, max_iterations=3000, g2=10, th_cv=1.2):
        self.target_path = target_path
        self.swarm_size = swarm_size
        self.max_iterations = max_iterations
        self.g2 = g2
        self.th_cv = th_cv

        # 完全随机初始化
        self.particles = [Particle() for _ in range(swarm_size)]
        self.global_best_position = None
        self.global_best_fitness = 0
        self.global_best_triggered = set()  # 核心修复：记录最高分对应的真实触发路径
        self.reset_count = 0

        # 初始化适应度评估
        for particle in self.particles:
            fit, trig = self.evaluate(particle.position)
            particle.fitness = fit
            particle.best_fitness = fit
            particle.best_position = particle.position.copy()

            if fit > self.global_best_fitness:
                self.global_best_fitness = fit
                self.global_best_position = particle.position.copy()
                self.global_best_triggered = trig

    def evaluate(self, position):
        """同时返回适应度和实际触发的路径"""
        try:
            triggered = execute_Tr(position)
            if self.target_path.issubset(triggered):
                return 1.0, triggered
            intersection = len(triggered & self.target_path)
            union = len(triggered | self.target_path)
            return (intersection / union if union > 0 else 0.0), triggered
        except:
            return 0.0, set()

    def calculate_cv(self, scores):
        """实现公式 25：综合变异系数"""
        N = self.swarm_size
        mean_fit = np.mean(scores)
        if mean_fit == 0: return float('inf')

        var_term = np.sum((scores - mean_fit) ** 2) / N
        adj_diff = sum(abs(scores[2 * i + 1] - scores[2 * i]) for i in range(N // 2))
        adj_term = ((2 / N) * adj_diff) ** 2

        numerator = math.sqrt(var_term + adj_term)
        denominator = 2 * np.sum(scores)

        if denominator == 0: return float('inf')
        return numerator / denominator

    def flip_mutation(self, position, rho=0.1):
        """实现公式 27：变异反转操作"""
        new_pos = np.copy(position)
        new_pos[0] = BOUNDS['light'][0] + BOUNDS['light'][1] - position[0] + rho
        new_pos[1] = BOUNDS['temp'][0] + BOUNDS['temp'][1] - position[1] + rho
        new_pos[2] = BOUNDS['moisture'][0] + BOUNDS['moisture'][1] - position[2] + rho
        return clip_state(new_pos)

    def optimize(self):
        w = 0.7
        c1 = 1.5
        c2 = 1.5

        max_velocity = np.array([
            (BOUNDS['light'][1] - BOUNDS['light'][0]) * 0.2,
            (BOUNDS['temp'][1] - BOUNDS['temp'][0]) * 0.2,
            (BOUNDS['moisture'][1] - BOUNDS['moisture'][0]) * 0.2
        ])

        c = 0
        early_stop = False
        converged_at = self.max_iterations

        for iteration in range(self.max_iterations):
            scores = np.array([p.fitness for p in self.particles])
            c += 1

            # MI-PSO 特有的停滞检测与变异 (对应 Algorithm 4 Line 20-23)
            if c == self.g2:
                cv = self.calculate_cv(scores)
                if cv <= self.th_cv:
                    self.reset_count += 1
                    l = int(round(self.swarm_size * (1 - cv)))
                    l = min(max(l, 1), self.swarm_size)

                    selected_indices = np.random.choice(self.swarm_size, l, replace=False)
                    for idx in selected_indices:
                        new_pos = self.flip_mutation(self.particles[idx].position)
                        new_fit, new_trig = self.evaluate(new_pos)

                        if new_fit > self.particles[idx].fitness:
                            self.particles[idx].position = new_pos
                            self.particles[idx].fitness = new_fit
                            if new_fit > self.particles[idx].best_fitness:
                                self.particles[idx].best_fitness = new_fit
                                self.particles[idx].best_position = new_pos.copy()
                            if new_fit > self.global_best_fitness:
                                self.global_best_fitness = new_fit
                                self.global_best_position = new_pos.copy()
                                self.global_best_triggered = new_trig  # 更新最优路径
                c = 0

                # 标准 PSO 速度与位置更新
            for particle in self.particles:
                r1, r2 = np.random.random(3), np.random.random(3)
                particle.velocity = (w * particle.velocity +
                                     c1 * r1 * (particle.best_position - particle.position) +
                                     c2 * r2 * (self.global_best_position - particle.position))
                particle.velocity = np.clip(particle.velocity, -max_velocity, max_velocity)

                particle.position += particle.velocity
                particle.position = clip_state(particle.position)

                fit, trig = self.evaluate(particle.position)
                particle.fitness = fit

                if fit > particle.best_fitness:
                    particle.best_fitness = fit
                    particle.best_position = particle.position.copy()
                if fit > self.global_best_fitness:
                    self.global_best_fitness = fit
                    self.global_best_position = particle.position.copy()
                    self.global_best_triggered = trig  # 更新最优路径

            if self.global_best_fitness >= 1.0:
                converged_at = iteration + 1
                early_stop = True
                break

        return self.global_best_position, self.global_best_fitness, self.global_best_triggered, converged_at, early_stop, self.reset_count


# === Excel导出函数 ===
def export_multiple_runs_to_excel(all_run_results, all_run_metrics, num_runs, filename=None):
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Isolated_MIPSO_{num_runs}Runs_{timestamp}.xlsx"

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    border = Border(
        left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 工作表1: 运行汇总
    ws1 = wb.active
    ws1.title = "运行汇总"
    ws1.sheet_view.showGridLines = False

    headers = ["运行次数", "成功率", "成功数量", "平均适应度", "平均迭代次数", "运行时间(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, border, center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, (results, run_metrics) in enumerate(zip(all_run_results, all_run_metrics), start=1):
        success_count = sum(1 for r in results if r['perfect_match'])
        success_rate = (success_count / len(targetPaths)) * 100
        avg_fitness = np.mean([r['fitness'] for r in results])

        iterations_list = [r['convergence_iteration'] if r.get('convergence_iteration') is not None else 10000 for r in
                           results]
        avg_iterations = np.mean(iterations_list)
        total_time = run_metrics.pso_end_time - run_metrics.pso_start_time if run_metrics.pso_end_time else 0

        row_data = [f"运行 {run_idx}", f"{success_rate:.1f}%", f"{success_count}/{len(targetPaths)}",
                    f"{avg_fitness:.4f}", f"{avg_iterations:.1f}", f"{total_time:.2f}"]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border, cell.alignment = border, center_align
            if run_idx % 2 == 0: cell.fill = alternate_fill
            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    # 工作表2: 路径统计
    ws2 = wb.create_sheet(title="路径统计")
    ws2.sheet_view.showGridLines = False
    headers2 = ["路径编号", "成功次数", "成功率", "平均适应度", "平均迭代次数", "最小迭代", "最大迭代"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, border, center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(targetPaths)):
        success_count = sum(1 for results in all_run_results if results[path_idx]['perfect_match'])
        success_rate = (success_count / num_runs) * 100
        avg_fitness = np.mean([results[path_idx]['fitness'] for results in all_run_results])

        iterations_list = [results[path_idx]['convergence_iteration'] if results[path_idx].get(
            'convergence_iteration') is not None else 10000 for results in all_run_results]
        avg_iterations = np.mean(iterations_list)

        row_data = [f"路径 {path_idx + 1}", f"{success_count}/{num_runs}", f"{success_rate:.1f}%",
                    f"{avg_fitness:.4f}", f"{avg_iterations:.1f}", np.min(iterations_list), np.max(iterations_list)]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border, cell.alignment = border, center_align
            if (path_idx + 1) % 2 == 0: cell.fill = alternate_fill
            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    # 工作表3: 最佳粒子详情
    ws3 = wb.create_sheet(title="最佳粒子详情")
    ws3.sheet_view.showGridLines = False
    headers3 = ["路径", "运行", "最佳粒子(light,temp,moisture)", "适应度", "迭代次数", "求解方法", "生成路径"]
    col_widths3 = [10, 10, 25, 12, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, border, center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(targetPaths)):
        for run_idx, results in enumerate(all_run_results, start=1):
            result = results[path_idx]
            particle_str = f"({int(result['best_position'][0])}, {int(result['best_position'][1])}, {int(result['best_position'][2])})"
            path_str = str(sorted(list(result['triggered'])))

            convergence_iter = result['convergence_iteration'] if result.get(
                'convergence_iteration') is not None else 10000

            row_data = [f"路径{path_idx + 1}", f"运行{run_idx}", particle_str, f"{result['fitness']:.4f}",
                        convergence_iter if convergence_iter < 10000 else "-", "MI-PSO", path_str]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = left_align if col == 7 else center_align

                if result['fitness'] == 1.0:
                    cell.fill = success_fill
                elif result['fitness'] < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill
            row_idx += 1

    # 保存
    output_dir = os.path.join(os.getcwd(), "results")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"\n✓ 独立 MI-PSO 消融实验结果已导出到: {filepath}")
    return filepath


def run_single_mipso_experiment(run_num):
    print(f"\n{'=' * 50}")
    print(f"开始第 {run_num} 次运行 (独立 MI-PSO)")
    print(f"{'=' * 50}")

    run_metrics = MetricsCollector()
    run_metrics.start_pso_phase()

    pso_results = []

    for i, target_path in enumerate(targetPaths):
        path_start_time = time.time()

        mipso = MIPSO(target_path, swarm_size=20, max_iterations=3000)
        # 直接接收并使用最佳寻优过程中真实触发的 best_trig，坚决不重新 evaluate
        best_pos, best_fit, best_trig, converged_at, early_stop, reset_count = mipso.optimize()

        path_execution_time = time.time() - path_start_time

        # 使用真实的最高分数判断是否完美
        is_perfect = (best_fit >= 1.0)

        pso_results.append({
            'target_path': target_path,
            'best_position': best_pos,
            'fitness': best_fit,
            'triggered': best_trig,
            'perfect_match': is_perfect,
            'method': 'MI-PSO',
            'convergence_iteration': converged_at,
            'early_stopped': early_stop,
            'reset_count': reset_count
        })

        run_metrics.record_pso_result(
            fitness=best_fit,
            is_perfect_match=is_perfect,
            convergence_iter=converged_at if early_stop else None,
            path_id=i + 1,
            method='MI-PSO',
            reset_count=reset_count,
            execution_time=path_execution_time
        )

        status = "✓完美(MI-PSO)" if is_perfect else f"○部分({best_fit:.3f})"
        print(f"  路径{i + 1}: {status} | 变异触发次数: {reset_count} | 耗时 {path_execution_time:.2f}s")

    run_metrics.end_pso_phase()

    success_count = sum(1 for r in pso_results if r['perfect_match'])
    success_rate = (success_count / len(targetPaths)) * 100
    pso_time = run_metrics.pso_end_time - run_metrics.pso_start_time

    print(f"\n第{run_num}次运行完成: 成功率 {success_rate:.1f}% | 耗时 {pso_time:.2f}秒")
    return pso_results, run_metrics


def run_multiple_mipso_experiments(num_runs):
    all_run_results = []
    all_run_metrics = []
    total_start = time.time()

    for run_num in range(1, num_runs + 1):
        results, metrics = run_single_mipso_experiment(run_num)
        all_run_results.append(results)
        all_run_metrics.append(metrics)

    total_time = time.time() - total_start
    print(f"\n{'=' * 50}\n全部{num_runs}次独立 MI-PSO 运行完成! 总耗时: {total_time:.2f}秒\n{'=' * 50}")

    return all_run_results, all_run_metrics


if __name__ == "__main__":
    if len(sys.argv) > 1:
        try:
            NUM_RUNS = int(sys.argv[1])
        except ValueError:
            pass

    print("=" * 70)
    print(" 独立 MI-PSO 消融实验 (无 DQN 介入，完美修复随机性 Bug)")
    print("=" * 70)

    all_run_results, all_run_metrics = run_multiple_mipso_experiments(NUM_RUNS)
    export_multiple_runs_to_excel(all_run_results, all_run_metrics, NUM_RUNS)