import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

def execute_validation_rules(a):
    """
    替换原有的编排规则函数
    参数a: 包含3个元素的元组或数组，分别对应path_depth, file_count, access_level
    返回: 触发的规则编号集合
    """
    path_depth, file_count, access_level = float(a[0]), int(a[1]), float(a[2])

    # 使用正确的变量名
    x, y, z = path_depth, file_count, access_level

    triggered = set()
    # 创建一个字典来存储b数组的值，用于跟踪哪些规则被触发
    b = {}
    y_score = max(0, 100 - y * 5)


    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((78 * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2): b[
    0] = 1
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((x * 95) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2): b[
    1] = 2
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((y * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2): b[
        2] = 3
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((z * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2): b[
        3] = 4
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((x * x) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2): b[
        4] = 5
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2): b[
        5] = 6
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((x * y) / (55 + 1) > 50 or x ** 2 + y ** 2 > z ** 2): b[
        6] = 7
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((x * y) / (x + 1) > 50 or x ** 2 + y ** 2 > z ** 2): b[
        7] = 8
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((x * y) / (y + 1) > 50 or x ** 2 + y ** 2 > z ** 2): b[
        8] = 9
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((x * y) / (20 + 1) > 50 or x ** 2 + y ** 2 > z ** 2): b[
        9] = 10
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 30 or x ** 2 + y ** 2 > z ** 2): b[
        10] = 11
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((x * y * 2) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2): b[
        11] = 12
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 33 or x ** 2 + y ** 2 > z ** 2): b[
        12] = 13
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2): b[
        13] = 14
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 or y ** 2 + y ** 2 > z ** 2): b[
        14] = 15
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 or 22 ** 2 + y ** 2 > z ** 2): b[
        15] = 16
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 or x ** 2 + x ** 2 > z ** 2): b[
        16] = 17
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 or x ** 2 + 55 ** 2 > z ** 2): b[
        17] = 18
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 or x ** 2.2 + y ** 2 > z ** 2): b[
        18] = 19
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 or x ** 2 + y ** 1.5 > z ** 2): b[
        19] = 20
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 != z ** 2): b[
        20] = 21
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2.3): b[
        21] = 22
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 1.4): b[
        22] = 23
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > 27 ** 2): b[
        23] = 24
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2.4 > z ** 2): b[
        24] = 25
    if ((x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 or x ** 1.4 + y ** 2 > z ** 2): b[
        25] = 26
    if (x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2:
        pattern_type = "221A1"
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[26] = 27
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (34 ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[27] = 28
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2.4 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[28] = 29
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 1.5 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[29] = 30
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - 35 ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[30] = 31
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - x ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[31] = 32
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2.7) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[32] = 33
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 1.2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[33] = 34
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2) / (x + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[34] = 35
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[35] = 36
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2) / (44 + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[36] = 37
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2) / (z + 3) < -30 or (abs(x - y) * z) / 100 > 45): b[37] = 38
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2) / (z + 0.1) != -30 or (abs(x - y) * z) / 100 > 45): b[38] = 39
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2) / (z + 0.1) < -6 or (abs(x - y) * z) / 100 > 45): b[39] = 40
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2) / (z + 0.1) < -30 and (abs(x - y) * z) / 100 > 45): b[40] = 41
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * z) / 100 > 45): b[41] = 42
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - z) * z) / 100 > 45): b[42] = 43
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(84 - y) * z) / 100 > 45): b[43] = 44
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - 93) * z) / 100 > 45): b[44] = 45
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * 69) / 100 > 45): b[45] = 46
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45): b[46] = 47
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * y) / 100 > 45): b[47] = 48
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 77 > 45): b[48] = 49
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 != 45): b[49] = 50
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 63): b[50] = 51
    if ((z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 33): b[51] = 52
    if (z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45:
        pattern_type = "221A2"
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((y + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35): b[52] = 53
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((z + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35): b[53] = 54
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((45 + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35): b[54] = 55
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((x + x) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35): b[55] = 56
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((x + z) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35): b[56] = 57
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((x + 65) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35): b[57] = 58
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2.5) ** 2 < y * 20 and (x * y * z) / 1000 < 35): b[58] = 59
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 1.3) ** 2 < y * 20 and (x * y * z) / 1000 < 35): b[59] = 60
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2.4 < y * 20 and (x * y * z) / 1000 < 35): b[60] = 61
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 != y * 20 and (x * y * z) / 1000 < 35): b[61] = 62
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < 33 * 20 and (x * y * z) / 1000 < 35): b[62] = 63
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < y * 15 and (x * y * z) / 1000 < 35): b[63] = 64
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < x * 20 and (x * y * z) / 1000 < 35): b[64] = 65
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35): b[65] = 66
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < y * 20 or (x * y * z) / 1000 < 35): b[66] = 67
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < y * 20 and (25 * y * z) / 1000 < 35): b[67] = 68
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < y * 20 and (x * 84 * z) / 1000 < 35): b[68] = 69
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < y * 20 and (x * y * 56) / 1000 < 35): b[69] = 70
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < y * 20 and (y * y * z) / 1000 < 35): b[70] = 71
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < y * 20 and (z * y * z) / 1000 < 35): b[71] = 72
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < y * 20 and (x * x * z) / 1000 < 35): b[72] = 73
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < y * 20 and (x * z * z) / 1000 < 35): b[73] = 74
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < y * 20 and (x * y * x) / 1000 < 35): b[74] = 75
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < y * 20 and (x * y * y) / 1000 < 35): b[75] = 76
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1070 < 35): b[76] = 77
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 != 35): b[77] = 78
    if (((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 47): b[78] = 79
    if ((x + y) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35:
        pattern_type = "221A3"
    if (x > 85 and z < 60) != (x > 85 or z < 60): b[79] = 80
    if (x > 85 and z < 60) != (x != 85 and z < 60): b[80] = 81
    if (x > 85 and z < 60) != (x > 94 and z < 60): b[81] = 82
    if (x > 85 and z < 60) != (x > 85 and z != 60): b[82] = 83
    if (x > 85 and z < 60) != (x > 85 and z < 42): b[83] = 84
    if x > 85 and z < 60:
        pattern_type = 222
    if ((z < 40 and x > 40) or (z < 45 and x > 45 and y_score < 50)) != (
            (z != 40 and x > 40) or (z < 45 and x > 45 and y_score < 50)): b[84] = 85
    if ((z < 40 and x > 40) or (z < 45 and x > 45 and y_score < 50)) != (
            (z < 49 and x > 40) or (z < 45 and x > 45 and y_score < 50)): b[85] = 86
    if ((z < 40 and x > 40) or (z < 45 and x > 45 and y_score < 50)) != (
            (z < 40 or x > 40) or (z < 45 and x > 45 and y_score < 50)): b[86] = 87
    if ((z < 40 and x > 40) or (z < 45 and x > 45 and y_score < 50)) != (
            (z < 40 and x != 40) or (z < 45 and x > 45 and y_score < 50)): b[87] = 88
    if ((z < 40 and x > 40) or (z < 45 and x > 45 and y_score < 50)) != (
            (z < 40 and x > 26) or (z < 45 and x > 45 and y_score < 50)): b[88] = 89
    if ((z < 40 and x > 40) or (z < 45 and x > 45 and y_score < 50)) != (
            (z < 40 and x > 40) and (z < 45 and x > 45 and y_score < 50)): b[89] = 90
    if ((z < 40 and x > 40) or (z < 45 and x > 45 and y_score < 50)) != (
            (z < 40 and x > 40) or (z != 45 and x > 45 and y_score < 50)): b[90] = 91
    if ((z < 40 and x > 40) or (z < 45 and x > 45 and y_score < 50)) != (
            (z < 40 and x > 40) or (z < 33 and x > 45 and y_score < 50)): b[91] = 92
    if ((z < 40 and x > 40) or (z < 45 and x > 45 and y_score < 50)) != (
            (z < 40 and x > 40) or (z < 45 or x > 45 and y_score < 50)): b[92] = 93
    if ((z < 40 and x > 40) or (z < 45 and x > 45 and y_score < 50)) != (
            (z < 40 and x > 40) or (z < 45 and x != 45 and y_score < 50)): b[93] = 94
    if ((z < 40 and x > 40) or (z < 45 and x > 45 and y_score < 50)) != (
            (z < 40 and x > 40) or (z < 45 and x > 66 and y_score < 50)): b[94] = 95
    if ((z < 40 and x > 40) or (z < 45 and x > 45 and y_score < 50)) != (
            (z < 40 and x > 40) or (z < 45 and x > 45 or y_score < 50)): b[95] = 96
    if ((z < 40 and x > 40) or (z < 45 and x > 45 and y_score < 50)) != (
            (z < 40 and x > 40) or (z < 45 and x > 45 and y_score != 50)): b[96] = 97
    if ((z < 40 and x > 40) or (z < 45 and x > 45 and y_score < 50)) != (
            (z < 40 and x > 40) or (z < 45 and x > 45 and y_score < 61)): b[97] = 98
    if (z < 40 and x > 40) or (z < 45 and x > 45 and y_score < 50):
        pattern_type = 223
    if (y_score < 30 and x > 30) != (y_score < 30 or x > 30): b[98] = 99
    if (y_score < 30 and x > 30) != (y_score < 30 and x != 30): b[99] = 100
    if (y_score < 30 and x > 30) != (y_score < 30 and x > 44): b[100] = 101
    if (y_score < 30 and x > 30) != (y_score != 30 and x > 30): b[101] = 102
    if (y_score < 30 and x > 30) != (y_score < 50 and x > 30): b[102] = 103
    if y_score < 30 and x > 30:
        pattern_type = 224
    if ((x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)) != (
            (x != 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)): b[103] = 104
    if ((x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)) != (
            (x > 56 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)): b[104] = 105
    if ((x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)) != (
            (x > 70 or z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)): b[105] = 106
    if ((x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)) != (
            (x > 70 and z < 70 and (x * (40 - z)) / 100 > 25) or (x > 75 and z < 65)): b[106] = 107
    if ((x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)) != (
            (x > 70 and z < 20 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)): b[107] = 108
    if ((x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)) != (
            (x > 70 and z < 70 or (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)): b[108] = 109
    if ((x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)) != (
            (x > 70 and z < 70 and (y * (80 - z)) / 100 > 25) or (x > 75 and z < 65)): b[109] = 110
    if ((x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)) != (
            (x > 70 and z < 70 and (z * (80 - z)) / 100 > 25) or (x > 75 and z < 65)): b[110] = 111
    if ((x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)) != (
            (x > 70 and z < 70 and (33 * (80 - z)) / 100 > 25) or (x > 75 and z < 65)): b[111] = 112
    if ((x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)) != (
            (x > 70 and z < 70 and (x * (78 - z)) / 100 > 25) or (x > 75 and z < 65)): b[112] = 113
    if ((x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)) != (
            (x > 70 and z < 70 and (x * (80 - x)) / 100 > 25) or (x > 75 and z < 65)): b[113] = 114
    if ((x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)) != (
            (x > 70 and z < 70 and (x * (80 - y)) / 100 > 25) or (x > 75 and z < 65)): b[114] = 115
    if ((x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)) != (
            (x > 70 and z < 70 and (x * (80 - 42)) / 100 > 25) or (x > 75 and z < 65)): b[115] = 116
    if ((x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)) != (
            (x > 70 and z < 70 and (x * (80 - z)) / 67 > 25) or (x > 75 and z < 65)): b[116] = 117
    if ((x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)) != (
            (x > 70 and z < 70 and (x * (80 - z)) / 100 != 25) or (x > 75 and z < 65)): b[117] = 118
    if ((x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)) != (
            (x > 70 and z < 70 and (x * (80 - z)) / 100 > 18) or (x > 75 and z < 65)): b[118] = 119
    if ((x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)) != (
            (x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) and (x > 75 and z < 65)): b[119] = 120
    if ((x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)) != (
            (x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x != 75 and z < 65)): b[120] = 121
    if ((x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)) != (
            (x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 88 and z < 65)): b[121] = 122
    if ((x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)) != (
            (x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 or z < 65)): b[122] = 123
    if ((x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)) != (
            (x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z != 65)): b[123] = 124
    if ((x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65)) != (
            (x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 47)): b[124] = 125
    if (x > 70 and z < 70 and (x * (80 - z)) / 100 > 25) or (x > 75 and z < 65):
        pattern_type = 225
    if (((z ** 2) / (x + y_score + 1) < 30) or (z < 65 and x > 50)) != (
            ((45 ** 2) / (x + y_score + 1) < 30) or (z < 65 and x > 50)): b[125] = 126
    if (((z ** 2) / (x + y_score + 1) < 30) or (z < 65 and x > 50)) != (
            ((x ** 2) / (x + y_score + 1) < 30) or (z < 65 and x > 50)): b[126] = 127
    if (((z ** 2) / (x + y_score + 1) < 30) or (z < 65 and x > 50)) != (
            ((y ** 2) / (x + y_score + 1) < 30) or (z < 65 and x > 50)): b[127] = 128
    if (((z ** 2) / (x + y_score + 1) < 30) or (z < 65 and x > 50)) != (
            ((z ** 2.5) / (x + y_score + 1) < 30) or (z < 65 and x > 50)): b[128] = 129
    if (((z ** 2) / (x + y_score + 1) < 30) or (z < 65 and x > 50)) != (
            ((z ** 2) / (y + y_score + 1) < 30) or (z < 65 and x > 50)): b[129] = 130
    if (((z ** 2) / (x + y_score + 1) < 30) or (z < 65 and x > 50)) != (
            ((z ** 2) / (z + y_score + 1) < 30) or (z < 65 and x > 50)): b[130] = 131
    if (((z ** 2) / (x + y_score + 1) < 30) or (z < 65 and x > 50)) != (
            ((z ** 2) / (66 + y_score + 1) < 30) or (z < 65 and x > 50)): b[131] = 132
    if (((z ** 2) / (x + y_score + 1) < 30) or (z < 65 and x > 50)) != (
            ((z ** 2) / (x + 25 + 1) < 30) or (z < 65 and x > 50)): b[132] = 133
    if (((z ** 2) / (x + y_score + 1) < 30) or (z < 65 and x > 50)) != (
            ((z ** 2) / (x + y_score + 1) != 30) or (z < 65 and x > 50)): b[133] = 134
    if (((z ** 2) / (x + y_score + 1) < 30) or (z < 65 and x > 50)) != (
            ((z ** 2) / (x + y_score + 1) < 44) or (z < 65 and x > 50)): b[134] = 135
    if (((z ** 2) / (x + y_score + 1) < 30) or (z < 65 and x > 50)) != (
            ((z ** 2) / (x + y_score + 1) < 30) and (z < 65 and x > 50)): b[135] = 136
    if (((z ** 2) / (x + y_score + 1) < 30) or (z < 65 and x > 50)) != (
            ((z ** 2) / (x + y_score + 1) < 30) or (z != 65 and x > 50)): b[136] = 137
    if (((z ** 2) / (x + y_score + 1) < 30) or (z < 65 and x > 50)) != (
            ((z ** 2) / (x + y_score + 1) < 30) or (z < 75 and x > 50)): b[137] = 138
    if (((z ** 2) / (x + y_score + 1) < 30) or (z < 65 and x > 50)) != (
            ((z ** 2) / (x + y_score + 1) < 30) or (z < 65 or x > 50)): b[138] = 139
    if (((z ** 2) / (x + y_score + 1) < 30) or (z < 65 and x > 50)) != (
            ((z ** 2) / (x + y_score + 1) < 30) or (z < 65 and x != 50)): b[139] = 140
    if (((z ** 2) / (x + y_score + 1) < 30) or (z < 65 and x > 50)) != (
            ((z ** 2) / (x + y_score + 1) < 30) or (z < 65 and x > 66)): b[140] = 141
    if (((z ** 2) / (x + y_score + 1) < 30) or (z < 65 and x > 50)) != (
            ((z ** 2) / (x + y_score + 10) < 30) or (z < 65 and x > 50)): b[141] = 142
    if ((z ** 2) / (x + y_score + 1) < 30) or (z < 65 and x > 50):
        pattern_type = 226
    if (40 <= y_score < 70 and z > 70) != (40 <= y_score < 70 or z > 70): b[142] = 143
    if (40 <= y_score < 70 and z > 70) != (40 <= y_score < 70 and z != 70): b[143] = 144
    if (40 <= y_score < 70 and z > 70) != (40 <= y_score < 70 and z > 81): b[144] = 145
    if (40 <= y_score < 70 and z > 70) != (40 <= y_score < 79 and z > 70): b[145] = 146
    if (40 <= y_score < 70 and z > 70) != (22 <= y_score < 70 and z > 70): b[146] = 147
    if 40 <= y_score < 70 and z > 70:
        pattern_type = 227
    if ((60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 65 and z > 80)) != (
            (60 <= x < 89 and z > 75 and (x * z) / 100 > 52) or (x > 65 and z > 80)): b[147] = 148
    if ((60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 65 and z > 80)) != (
            (46 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 65 and z > 80)): b[148] = 149
    if ((60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 65 and z > 80)) != (
            (60 <= x < 80 or z > 75 and (x * z) / 100 > 52) or (x > 65 and z > 80)): b[149] = 150
    if ((60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 65 and z > 80)) != (
            (60 <= x < 80 and z != 75 and (x * z) / 100 > 52) or (x > 65 and z > 80)): b[150] = 151
    if ((60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 65 and z > 80)) != (
            (60 <= x < 80 and z > 66 and (x * z) / 100 > 52) or (x > 65 and z > 80)): b[151] = 152
    if ((60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 65 and z > 80)) != (
            (60 <= x < 80 and z > 75 or (x * z) / 100 > 52) or (x > 65 and z > 80)): b[152] = 153
    if ((60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 65 and z > 80)) != (
            (60 <= x < 80 and z > 75 and (z * z) / 100 > 52) or (x > 65 and z > 80)): b[153] = 154
    if ((60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 65 and z > 80)) != (
            (60 <= x < 80 and z > 75 and (x * x) / 100 > 52) or (x > 65 and z > 80)): b[154] = 155
    if ((60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 65 and z > 80)) != (
            (60 <= x < 80 and z > 75 and (55 * z) / 100 > 52) or (x > 65 and z > 80)): b[155] = 156
    if ((60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 65 and z > 80)) != (
            (60 <= x < 80 and z > 75 and (x * 45) / 100 > 52) or (x > 65 and z > 80)): b[156] = 157
    if ((60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 65 and z > 80)) != (
            (60 <= x < 80 and z > 75 and (x * z) / 92 > 52) or (x > 65 and z > 80)): b[157] = 158
    if ((60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 65 and z > 80)) != (
            (60 <= x < 80 and z > 75 and (x * z) / 100 != 52) or (x > 65 and z > 80)): b[158] = 159
    if ((60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 65 and z > 80)) != (
            (60 <= x < 80 and z > 75 and (x * z) / 100 > 32) or (x > 65 and z > 80)): b[159] = 160
    if ((60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 65 and z > 80)) != (
            (60 <= x < 80 and z > 75 and (x * z) / 100 > 52) and (x > 65 and z > 80)): b[160] = 161
    if ((60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 65 and z > 80)) != (
            (60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x != 65 and z > 80)): b[161] = 162
    if ((60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 65 and z > 80)) != (
            (60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 24 and z > 80)): b[162] = 163
    if ((60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 65 and z > 80)) != (
            (60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 65 or z > 80)): b[163] = 164
    if ((60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 65 and z > 80)) != (
            (60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 65 and z != 80)): b[164] = 165
    if ((60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 65 and z > 80)) != (
            (60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 65 and z > 91)): b[165] = 166
    if (60 <= x < 80 and z > 75 and (x * z) / 100 > 52) or (x > 65 and z > 80):
        pattern_type = 228
    if ((x < 30 and 70 <= y_score < 85) or (x < 35 and 75 <= y_score < 90)) != (
            (x != 30 and 70 <= y_score < 85) or (x < 35 and 75 <= y_score < 90)): b[166] = 167
    if ((x < 30 and 70 <= y_score < 85) or (x < 35 and 75 <= y_score < 90)) != (
            (x < 10 and 70 <= y_score < 85) or (x < 35 and 75 <= y_score < 90)): b[167] = 168
    if ((x < 30 and 70 <= y_score < 85) or (x < 35 and 75 <= y_score < 90)) != (
            (x < 30 or 70 <= y_score < 85) or (x < 35 and 75 <= y_score < 90)): b[168] = 169
    if ((x < 30 and 70 <= y_score < 85) or (x < 35 and 75 <= y_score < 90)) != (
            (x < 30 and 55 <= y_score < 85) or (x < 35 and 75 <= y_score < 90)): b[169] = 170
    if ((x < 30 and 70 <= y_score < 85) or (x < 35 and 75 <= y_score < 90)) != (
            (x < 30 and 70 <= y_score < 93) or (x < 35 and 75 <= y_score < 90)): b[170] = 171
    if ((x < 30 and 70 <= y_score < 85) or (x < 35 and 75 <= y_score < 90)) != (
            (x < 30 and 70 <= y_score < 85) and (x < 35 and 75 <= y_score < 90)): b[171] = 172
    if ((x < 30 and 70 <= y_score < 85) or (x < 35 and 75 <= y_score < 90)) != (
            (x < 30 and 70 <= y_score < 85) or (x != 35 and 75 <= y_score < 90)): b[172] = 173
    if ((x < 30 and 70 <= y_score < 85) or (x < 35 and 75 <= y_score < 90)) != (
            (x < 30 and 70 <= y_score < 85) or (x < 52 and 75 <= y_score < 90)): b[173] = 174
    if ((x < 30 and 70 <= y_score < 85) or (x < 35 and 75 <= y_score < 90)) != (
            (x < 30 and 70 <= y_score < 85) or (x < 35 or 75 <= y_score < 90)): b[174] = 175
    if ((x < 30 and 70 <= y_score < 85) or (x < 35 and 75 <= y_score < 90)) != (
            (x < 30 and 70 <= y_score < 85) or (x < 35 and 66 <= y_score < 90)): b[175] = 176
    if ((x < 30 and 70 <= y_score < 85) or (x < 35 and 75 <= y_score < 90)) != (
            (x < 30 and 70 <= y_score < 85) or (x < 35 and 75 <= y_score < 96)): b[176] = 177
    if (x < 30 and 70 <= y_score < 85) or (x < 35 and 75 <= y_score < 90):
        pattern_type = 229
    if (30 <= x <= 60 and y_score >= 85 and z >= 85) != (30 <= x <= 69 and y_score >= 85 and z >= 85): b[177] = 178
    if (30 <= x <= 60 and y_score >= 85 and z >= 85) != (41 <= x <= 60 and y_score >= 85 and z >= 85): b[178] = 179
    if (30 <= x <= 60 and y_score >= 85 and z >= 85) != (30 <= x <= 60 or y_score >= 85 and z >= 85): b[179] = 180
    if (30 <= x <= 60 and y_score >= 85 and z >= 85) != (30 <= x <= 60 and y_score != 85 and z >= 85): b[180] = 181
    if (30 <= x <= 60 and y_score >= 85 and z >= 85) != (30 <= x <= 60 and y_score >= 75 and z >= 85): b[181] = 182
    if (30 <= x <= 60 and y_score >= 85 and z >= 85) != (30 <= x <= 60 and y_score >= 85 or z >= 85): b[182] = 183
    if (30 <= x <= 60 and y_score >= 85 and z >= 85) != (30 <= x <= 60 and y_score >= 85 and z != 85): b[183] = 184
    if (30 <= x <= 60 and y_score >= 85 and z >= 85) != (30 <= x <= 60 and y_score >= 85 and z >= 51): b[184] = 185
    if 30 <= x <= 60 and y_score >= 85 and z >= 85:
        pattern_type = 230

    # 返回被触发的规则编号集合
    return set(b.values())


def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """"""
    generated_path = execute_validation_rules(particle[0], particle[1], particle[2])

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # : x:1-50, y:1-50, z:1-50
        self.bounds = bounds if bounds else [(2, 100), (2, 100), (2, 100)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """target pathsPSO"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(particles[i][0], particles[i][1], particles[i][2]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(gbest_particle[0], gbest_particle[1], gbest_particle[2]),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """Path PSO"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - Path ")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations")
    print(f"Path : {len(target_paths)}")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"Path {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "" if result['success'] else f"({result['best_fitness']:.3f})"
        print(f"{status} | {result['time']:.2f}s | iterations{result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f": {success_count}/{len(target_paths)} ({success_rate:.1f}%) | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs}")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations, {len(target_paths)}Path ")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"---  {run_idx}/{num_runs} ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f": {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs} runcompleted | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 1: 
    ws1 = wb.active
    ws1.title = ""
    ws1.sheet_view.showGridLines = False

    headers = ["Run", "", "", "", "Average Iterations", "(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f" {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # 2: Path 
    ws2 = wb.create_sheet(title="Path ")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Path ID", "", "", "", "Average Iterations", "Minimum Iterations", "Maximum Iterations"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"Path  {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # 3: 
    ws3 = wb.create_sheet(title="")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Path ", "", "(x,y,z)", "", "Iterations", "Path "]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"{run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # 4: target paths
    ws4 = wb.create_sheet(title="target paths")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Path ID", "target paths", ""]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"Path  {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f" : {filename}")
    print(f"{'=' * 70}")
    print(f":")
    print(f"  1.        - {len(all_results)} run")
    print(f"  2. Path        - Path ")
    print(f"  3.    -  runPath ")
    print(f"  4. target paths       - target paths")
    print(f"{'=' * 70}\n")

    return filename


def main():


    target_paths = [
        {2, 5, 6, 9, 17, 18, 19, 21, 23, 24, 41, 43, 45, 46, 47, 48, 51, 53, 62, 65, 66, 67, 85, 87, 99, 102, 126, 128,
         134, 135, 137, 139, 143, 146, 150, 153, 154, 158, 159, 160, 164, 165, 167, 169, 173, 175},
        {14, 17, 20, 28, 29, 31, 32, 34, 36, 37, 38, 41, 53, 54, 55, 58, 60, 61, 63, 64, 65, 66, 71, 80, 81, 87, 88, 93,
         94, 96, 99, 100, 121, 123, 127, 130, 131, 132, 133, 134, 135, 139, 140, 142, 169, 175},
        {2, 6, 9, 17, 18, 19, 21, 23, 24, 41, 43, 45, 46, 47, 48, 51, 53, 62, 65, 66, 67, 85, 87, 99, 102, 126, 127,
         128, 134, 137, 139, 143, 150, 153, 154, 158, 159, 160, 162, 163, 164, 167, 169, 173, 175},
        {14, 17, 20, 22, 29, 31, 32, 34, 36, 37, 41, 53, 54, 55, 58, 60, 61, 63, 64, 65, 66, 68, 69, 70, 71, 72, 80, 81,
         87, 88, 93, 94, 96, 99, 100, 121, 123, 127, 130, 132, 133, 134, 139, 140, 169, 175},
        {2, 6, 21, 23, 24, 41, 43, 45, 46, 47, 48, 51, 53, 62, 65, 66, 67, 85, 87, 99, 102, 126, 127, 128, 134, 137,
         139, 143, 146, 149, 150, 153, 162, 163, 164, 167, 169, 173, 175, 180, 181, 182, 183},
        {14, 17, 20, 22, 27, 30, 33, 35, 39, 40, 50, 53, 54, 55, 58, 60, 61, 63, 64, 65, 66, 71, 72, 80, 81, 87, 88, 93,
         94, 96, 99, 100, 121, 123, 127, 130, 131, 132, 133, 134, 139, 140, 169, 175},
        {2, 6, 19, 21, 23, 24, 41, 43, 45, 46, 47, 48, 51, 53, 62, 65, 66, 67, 85, 87, 99, 102, 126, 127, 128, 134, 137,
         139, 143, 146, 155, 156, 157, 161, 167, 169, 173, 175, 180, 181, 182, 183},
        {1, 2, 16, 18, 19, 21, 23, 25, 33, 39, 50, 54, 55, 57, 58, 60, 61, 64, 69, 80, 81, 87, 88, 93, 94, 96, 99, 100,
         121, 123, 127, 128, 130, 131, 132, 133, 134, 135, 139, 140, 142, 169, 175},
        {14, 15, 16, 22, 26, 41, 42, 43, 45, 48, 51, 53, 62, 67, 85, 87, 99, 102, 106, 109, 115, 116, 118, 123, 124,
         127, 129, 130, 131, 132, 133, 136, 150, 151, 152, 153, 164, 165, 167, 169},
        {14, 17, 18, 20, 22, 27, 28, 30, 33, 35, 37, 39, 40, 45, 46, 49, 50, 52, 53, 54, 55, 60, 61, 63, 64, 65, 66, 68,
         71, 72, 80, 81, 96, 99, 100, 121, 123, 127, 134, 139, 140, 169, 175},
        {2, 5, 6, 9, 17, 18, 19, 21, 23, 24, 39, 42, 44, 49, 50, 52, 53, 62, 65, 66, 67, 85, 87, 99, 102, 126, 128, 134,
         135, 137, 138, 139, 143, 146, 150, 164, 165, 167, 169, 173, 175},
        {14, 15, 16, 22, 26, 41, 42, 43, 45, 48, 51, 53, 62, 67, 80, 83, 85, 87, 99, 102, 106, 109, 115, 116, 118, 123,
         124, 127, 129, 130, 133, 136, 153, 164, 165, 167, 169, 173, 175},
        {14, 15, 16, 22, 26, 41, 43, 45, 48, 53, 62, 67, 80, 83, 85, 87, 99, 102, 106, 123, 124, 126, 128, 134, 135,
         137, 139, 143, 146, 148, 150, 153, 164, 165, 167, 169, 173, 175},
        {14, 20, 28, 29, 31, 32, 34, 36, 37, 38, 41, 56, 57, 59, 62, 67, 80, 81, 87, 88, 93, 94, 96, 99, 100, 121, 123,
         127, 130, 131, 132, 133, 134, 135, 139, 140, 142, 169, 175},
        {14, 15, 16, 22, 26, 41, 42, 43, 45, 48, 51, 80, 83, 85, 87, 91, 93, 96, 99, 102, 103, 106, 109, 115, 116, 118,
         123, 124, 126, 128, 134, 135, 137, 138, 139, 153, 164, 165},
        {1, 2, 3, 4, 6, 7, 8, 9, 10, 11, 12, 13, 15, 19, 21, 23, 24, 25, 28, 30, 33, 39, 50, 85, 87, 91, 93, 96, 134,
         137, 139, 143, 149, 150, 153, 162, 163, 164, 180, 181, 183},
        {1, 15, 18, 21, 23, 24, 25, 27, 33, 39, 50, 53, 54, 55, 57, 58, 60, 61, 63, 64, 65, 68, 69, 70, 71, 72, 74, 80,
         81, 96, 99, 100, 121, 123, 127, 134, 139, 140, 169, 175},
        {14, 15, 16, 20, 22, 26, 33, 39, 50, 53, 62, 63, 65, 66, 67, 80, 81, 87, 88, 89, 93, 94, 96, 101, 121, 123, 127,
         128, 131, 132, 133, 134, 135, 139, 140, 142, 175, 180},
        {2, 6, 17, 18, 19, 21, 23, 24, 39, 42, 44, 49, 50, 52, 53, 62, 65, 66, 67, 85, 87, 99, 102, 126, 127, 128, 134,
         135, 137, 138, 139, 143, 173, 175, 180, 183, 184, 185},
        {14, 15, 16, 22, 26, 32, 39, 46, 47, 49, 50, 52, 53, 62, 67, 80, 81, 85, 87, 99, 102, 106, 109, 115, 116, 117,
         118, 119, 121, 123, 150, 164, 165, 167, 169, 173, 175},
        {2, 21, 23, 24, 39, 42, 44, 49, 50, 52, 53, 62, 63, 65, 66, 67, 85, 87, 99, 102, 126, 127, 128, 134, 143, 162,
         163, 164, 167, 169, 173, 174, 175, 180, 181, 182, 183},
        {2, 5, 6, 9, 17, 18, 19, 21, 23, 24, 41, 43, 45, 48, 53, 62, 67, 85, 87, 99, 102, 106, 123, 124, 126, 128, 134,
         135, 137, 139, 143, 161, 166, 173, 175, 180, 183},
        {2, 6, 9, 17, 18, 19, 21, 23, 24, 39, 42, 44, 49, 50, 52, 85, 87, 99, 102, 126, 127, 128, 134, 137, 139, 145,
         150, 153, 154, 158, 159, 160, 162, 163, 164, 180},
        {2, 6, 21, 23, 24, 41, 43, 45, 46, 47, 48, 51, 53, 62, 65, 66, 67, 85, 87, 99, 102, 126, 127, 128, 134, 137,
         139, 143, 155, 157, 161, 173, 175, 178, 180, 183},
        {14, 15, 16, 22, 26, 32, 39, 47, 50, 53, 62, 67, 80, 81, 85, 86, 87, 93, 96, 97, 98, 99, 102, 107, 108, 110,
         111, 112, 113, 114, 120, 143, 144, 150, 164, 165},
        {14, 15, 16, 22, 26, 41, 42, 43, 45, 48, 51, 85, 87, 91, 93, 96, 99, 102, 103, 106, 123, 124, 126, 128, 134,
         135, 137, 138, 139, 143, 147, 153, 164, 165},
        {14, 15, 16, 26, 32, 39, 46, 47, 50, 53, 62, 65, 67, 80, 81, 90, 99, 102, 107, 108, 110, 111, 112, 114, 120,
         150, 164, 165, 167, 169, 173, 175},
        {27, 28, 34, 35, 36, 37, 38, 41, 53, 54, 62, 63, 65, 67, 80, 81, 87, 88, 89, 93, 94, 96, 99, 102, 103, 104, 106,
         109, 121, 123, 136, 175, 180},
        {21, 23, 39, 42, 43, 44, 45, 50, 54, 55, 56, 57, 58, 60, 61, 64, 68, 69, 72, 73, 74, 96, 99, 100, 127, 128, 134,
         143, 162, 164, 169, 175, 183},
        {18, 21, 23, 39, 50, 54, 55, 57, 58, 60, 61, 69, 80, 81, 121, 123, 126, 127, 128, 131, 132, 134, 135, 139, 140,
         142, 143, 144, 169, 170, 175},
        {1, 15, 16, 18, 21, 23, 24, 25, 27, 30, 33, 39, 50, 67, 73, 75, 76, 77, 78, 79, 80, 81, 96, 99, 100, 121, 123,
         127, 134, 139, 140, 169, 175},
        {2, 5, 6, 17, 18, 19, 21, 23, 24, 39, 44, 49, 50, 52, 53, 62, 65, 66, 67, 85, 87, 99, 102, 121, 123, 136, 141,
         150, 167, 169, 173, 175, 180},
        {21, 39, 44, 45, 50, 53, 59, 62, 63, 65, 66, 67, 80, 81, 121, 123, 126, 127, 128, 131, 132, 134, 135, 139, 140,
         142, 169, 171, 175, 177},
        {14, 15, 16, 22, 26, 32, 39, 50, 53, 62, 67, 80, 81, 86, 87, 93, 96, 98, 99, 102, 104, 105, 106, 109, 121, 123,
         143, 144, 150, 164, 165},
        {21, 23, 24, 39, 42, 43, 44, 45, 50, 53, 62, 63, 65, 66, 67, 99, 102, 126, 127, 128, 134, 143, 162, 163, 164,
         172, 180, 181, 182, 183},
        {14, 15, 16, 22, 26, 32, 39, 44, 46, 47, 49, 50, 52, 85, 87, 91, 93, 96, 99, 102, 103, 120, 122, 125, 136, 143,
         144, 150, 164, 165},
        {21, 23, 24, 39, 42, 43, 44, 45, 50, 53, 62, 63, 65, 66, 67, 99, 102, 126, 127, 128, 134, 143, 162, 163, 164,
         174, 175, 179, 181},
        {14, 15, 16, 26, 32, 39, 46, 47, 49, 50, 52, 53, 62, 67, 82, 84, 85, 86, 87, 93, 96, 97, 99, 102, 164, 165, 167,
         169, 173, 175},
        {18, 21, 39, 50, 54, 55, 57, 58, 60, 61, 69, 80, 81, 121, 123, 126, 127, 128, 131, 132, 134, 135, 139, 140, 142,
         168, 172},
        {14, 15, 16, 26, 32, 39, 50, 80, 81, 90, 92, 99, 102, 103, 107, 108, 110, 111, 112, 114, 120, 150, 164, 165},
        {14, 15, 16, 22, 26, 39, 50, 80, 81, 90, 92, 95, 121, 123, 128, 132, 133, 134, 135, 139, 140, 142, 180},
        {2, 17, 18, 19, 21, 23, 24, 39, 50, 53, 62, 63, 65, 66, 67, 80, 81, 87, 88, 89, 93, 99, 102, 121, 123, 129, 136,
         167, 169, 175, 176, 180}
    ]

    print("=" * 70)
    print("baseline PSO")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("Program completed")


if __name__ == "__main__":
    main()