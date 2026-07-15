import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

def execute_validation_rules(a):
    """
    替换原有的编排规则函数
    参数a: 包含3个元素的元组或数组，分别对应path_depth, file_count, access_level
    返回: 触发的规则编号集合
    """
    path_depth, file_count, access_level = float(a[0]), int(a[1]), float(a[2])

    # 使用正确的变量名
    x, y, z = path_depth, file_count, access_level

    triggered = set()
    # 创建一个字典来存储b数组的值，用于跟踪哪些规则被触发
    b = {}
    x_score = max(0, 100 - x * 1.67)

    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[0] = 1
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((z * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[1] = 2
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * x) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[2] = 3
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[3] = 4
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * 43) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[4] = 5
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((74 * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[5] = 6
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (47 + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[6] = 7
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (x + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[7] = 8
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (y + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[8] = 9
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 10) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[9] = 10
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) != 50 and x ** 2 + y ** 2 > z ** 2):
        b[10] = 11
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 29 and x ** 2 + y ** 2 > z ** 2):
        b[11] = 12
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 69 and x ** 2 + y ** 2 > z ** 2):
        b[12] = 13
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2): b[
        13] = 14
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and y ** 2 + y ** 2 > z ** 2):
        b[14] = 15
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 3 + y ** 2 > z ** 2):
        b[15] = 16
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + x ** 2 > z ** 2):
        b[16] = 17
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 3 > z ** 2):
        b[17] = 18
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 != z ** 2):
        b[18] = 19
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2.5):
        b[19] = 20
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 1.4 > z ** 2):
        b[20] = 21
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + 34 ** 2 > z ** 2):
        b[21] = 22
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and 64 ** 2 + y ** 2 > z ** 2):
        b[22] = 23
    if (x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2:
        pattern_type = "311A1"
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[23] = 24
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - z ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[24] = 25
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 1.5) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[25] = 26
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2.3 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[26] = 27
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (45 ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[27] = 28
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - 31 ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[28] = 29
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[29] = 30
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (x + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[30] = 31
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 2) < -30 or (abs(x - y) * z) / 100 > 45): b[31] = 32
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (26 + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[32] = 33
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y - 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[33] = 34
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) != -30 or (abs(x - y) * z) / 100 > 45): b[34] = 35
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -10 or (abs(x - y) * z) / 100 > 45): b[35] = 36
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 and (abs(x - y) * z) / 100 > 45): b[36] = 37
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(z - y) * z) / 100 > 45): b[37] = 38
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - z) * z) / 100 > 45): b[38] = 39
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * y) / 100 > 45): b[39] = 40
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * x) / 100 > 45): b[40] = 41
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * 34) / 100 > 45): b[41] = 42
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - 26) * z) / 100 > 45): b[42] = 43
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(52 - y) * z) / 100 > 45): b[43] = 44
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 76 > 45): b[44] = 45
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 != 45): b[45] = 46
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 55): b[46] = 47
    if (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45:
        pattern_type = "311A2"

    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + x) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35): b[47] = 48
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((y + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35): b[48] = 49
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((24 + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35): b[49] = 50
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + 31) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35): b[50] = 51
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 1.5) ** 2 < z * 20 and (x * x * z) / 1000 < 35): b[51] = 52
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2.4 < z * 20 and (x * x * z) / 1000 < 35): b[52] = 53
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 != z * 20 and (x * x * z) / 1000 < 35): b[53] = 54
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < 25 * 20 and (x * x * z) / 1000 < 35): b[54] = 55
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 14 and (x * x * z) / 1000 < 35): b[55] = 56
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < x * 20 and (x * x * z) / 1000 < 35): b[56] = 57
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < y * 20 and (x * x * z) / 1000 < 35): b[57] = 58
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 1.2 < z * 20 and (x * x * z) / 1000 < 35): b[58] = 59
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 or (x * x * z) / 1000 < 35): b[59] = 60
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (y * x * z) / 1000 < 35): b[60] = 61
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (z * x * z) / 1000 < 35): b[61] = 62
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * 35 * z) / 1000 < 35): b[62] = 63
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * x * x) / 1000 < 35): b[63] = 64
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * x * y) / 1000 < 35): b[64] = 65
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * x * 26) / 1000 < 35): b[65] = 66
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * 54 * z) / 1000 < 35): b[66] = 67
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1050 < 35): b[67] = 68
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 != 35): b[68] = 69
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 25): b[69] = 70
    if ((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35:
        pattern_type = "311A3"
    if (x < 40 and z < 50) != (x != 40 and z < 50): b[70] = 71
    if (x < 40 and z < 50) != (x < 29 and z < 50): b[71] = 72
    if (x < 40 and z < 50) != (x < 40 or z < 50): b[72] = 73
    if (x < 40 and z < 50) != (x < 40 and z != 50): b[73] = 74
    if (x < 40 and z < 50) != (x < 40 and z < 36): b[74] = 75
    if x < 40 and z < 50:
        pattern_type = 322
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 or x > 50) or (z < 40 and x > 55 and y < 40)): b[75] = 76
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z != 35 and x > 50) or (z < 40 and x > 55 and y < 40)): b[76] = 77
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 45 and x > 50) or (z < 40 and x > 55 and y < 40)): b[77] = 78
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x != 50) or (z < 40 and x > 55 and y < 40)): b[78] = 79
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 59) or (z < 40 and x > 55 and y < 40)): b[79] = 80
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) and (z < 40 and x > 55 and y < 40)): b[80] = 81
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) or (z != 40 and x > 55 and y < 40)): b[81] = 82
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) or (z < 48 and x > 55 and y < 40)): b[82] = 83
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) or (z < 40 or x > 55 and y < 40)): b[83] = 84
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) or (z < 40 and x != 55 and y < 40)): b[84] = 85
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) or (z < 40 and x > 65 and y < 40)): b[85] = 86
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) or (z < 40 and x > 55 or y < 40)): b[86] = 87
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) or (z < 40 and x > 55 and y != 40)): b[87] = 88
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) or (z < 40 and x > 55 and y < 29)): b[88] = 89
    if (z < 35 and x > 50) or (z < 40 and x > 55 and y < 40):
        pattern_type = 323
    if (y < 25 and x > 60) != (y < 25 or x > 60): b[89] = 90
    if (y < 25 and x > 60) != (y != 25 and x > 60): b[90] = 91
    if (y < 25 and x > 60) != (y < 17 and x > 60): b[91] = 92
    if (y < 25 and x > 60) != (y < 25 and x != 60): b[92] = 93
    if (y < 25 and x > 60) != (y < 25 and x > 69): b[93] = 94
    if y < 25 and x > 60:
        pattern_type = 324
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 75 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)): b[94] = 95
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (22 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)): b[95] = 96
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 or 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)): b[96] = 97
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 29 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)): b[97] = 98
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 78 and (x * z) / 100 < 35) or (x < 60 and z < 60)): b[98] = 99
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 or (x * z) / 100 < 35) or (x < 60 and z < 60)): b[99] = 100
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (56 * z) / 100 < 35) or (x < 60 and z < 60)): b[100] = 101
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * 74) / 100 < 35) or (x < 60 and z < 60)): b[101] = 102
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 89 < 35) or (x < 60 and z < 60)): b[102] = 103
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 != 35) or (x < 60 and z < 60)): b[103] = 104
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 21) or (x < 60 and z < 60)): b[104] = 105
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (z * z) / 100 < 35) or (x < 60 and z < 60)): b[105] = 106
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * x) / 100 < 35) or (x < 60 and z < 60)): b[106] = 107
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) and (x < 60 and z < 60)): b[107] = 108
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x != 60 and z < 60)): b[108] = 109
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 or z < 60)): b[109] = 110
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z != 60)): b[110] = 111
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 68)): b[111] = 112
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 51 and z < 60)): b[112] = 113
    if (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60):
        pattern_type = 325
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((53 ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)): b[113] = 114
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 2.3) / (x + y + 1) < 20) or (z < 65 and x < 75)): b[114] = 115
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 1.4) / (x + y + 1) < 20) or (z < 65 and x < 75)): b[115] = 116
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((x ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)):
        b[116] = 117
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((y ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)):
        b[117] = 118
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((z ** 2) / (y + y + 1) < 20) or (z < 65 and x < 75)):
        b[118] = 119
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((z ** 2) / (z + y + 1) < 20) or (z < 65 and x < 75)):
        b[119] = 120
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((z ** 2) / (x + x + 1) < 20) or (z < 65 and x < 75)):
        b[120] = 121
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((z ** 2) / (x + z + 1) < 20) or (z < 65 and x < 75)):
        b[121] = 122
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 2) / (x + 45 + 1) < 20) or (z < 65 and x < 75)): b[122] = 123
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 2) / (33 + y + 1) < 20) or (z < 65 and x < 75)): b[123] = 124
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 2) / (x + y + 11) < 20) or (z < 65 and x < 75)): b[124] = 125
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 2) / (x + y + 1) != 20) or (z < 65 and x < 75)): b[125] = 126
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((z ** 2) / (x + y + 1) < 15) or (z < 65 and x < 75)):
        b[126] = 127
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 2) / (x + y + 1) < 20) and (z < 65 and x < 75)): b[127] = 128
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 2) / (x + y + 1) < 20) or (z != 65 and x < 75)): b[128] = 129
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((z ** 2) / (x + y + 1) < 20) or (z < 75 and x < 75)):
        b[129] = 130
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((z ** 2) / (x + y + 1) < 20) or (z < 65 or x < 75)):
        b[130] = 131
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 2) / (x + y + 1) < 20) or (z < 65 and x != 75)): b[131] = 132
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 79)):
        b[132] = 133
    if ((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75):
        pattern_type = 326
    if (60 <= x < 82 and y > 50) != (60 <= x < 82 or y > 50): b[133] = 134
    if (60 <= x < 82 and y > 50) != (60 <= x < 89 and y > 50): b[134] = 135
    if (60 <= x < 82 and y > 50) != (67 <= x < 82 and y > 50): b[135] = 136
    if (60 <= x < 82 and y > 50) != (60 <= x < 82 and y != 50): b[136] = 137
    if (60 <= x < 82 and y > 50) != (60 <= x < 82 and y > 62): b[137] = 138
    if 60 <= x < 82 and y > 50:
        pattern_type = 327
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 77 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)): b[138] = 139
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (22 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)): b[139] = 140
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 or x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)): b[140] = 141
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x != 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)): b[141] = 142
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 85 and (y * x) / 100 < 52) or (y < 60 and x > 80)): b[142] = 143
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 or (y * x) / 100 < 52) or (y < 60 and x > 80)): b[143] = 144
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (x * x) / 100 < 52) or (y < 60 and x > 80)): b[144] = 145
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * y) / 100 < 52) or (y < 60 and x > 80)): b[145] = 146
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (43 * x) / 100 < 52) or (y < 60 and x > 80)): b[146] = 147
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * 31) / 100 < 52) or (y < 60 and x > 80)): b[147] = 148
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 78 < 52) or (y < 60 and x > 80)): b[148] = 149
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 100 != 52) or (y < 60 and x > 80)): b[149] = 150
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 100 < 62) or (y < 60 and x > 80)): b[150] = 151
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 100 < 52) and (y < 60 and x > 80)): b[151] = 152
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y != 60 and x > 80)): b[152] = 153
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 65 and x > 80)): b[153] = 154
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 or x > 80)): b[154] = 155
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x != 80)): b[155] = 156
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 88)): b[156] = 157
    if (40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80):
        pattern_type = 328

    # 返回被触发的规则编号集合
    return set(b.values())

def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """"""
    generated_path = execute_validation_rules(particle[0], particle[1], particle[2])

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # : x:1-50, y:1-50, z:1-50
        self.bounds = bounds if bounds else [(2, 100), (2, 100), (2, 100)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """target pathsPSO"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(particles[i][0], particles[i][1], particles[i][2]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(gbest_particle[0], gbest_particle[1], gbest_particle[2]),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """Path PSO"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - Path ")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations")
    print(f"Path : {len(target_paths)}")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"Path {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "" if result['success'] else f"({result['best_fitness']:.3f})"
        print(f"{status} | {result['time']:.2f}s | iterations{result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f": {success_count}/{len(target_paths)} ({success_rate:.1f}%) | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs}")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations, {len(target_paths)}Path ")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"---  {run_idx}/{num_runs} ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f": {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs} runcompleted | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 1: 
    ws1 = wb.active
    ws1.title = ""
    ws1.sheet_view.showGridLines = False

    headers = ["Run", "", "", "", "Average Iterations", "(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f" {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # 2: Path 
    ws2 = wb.create_sheet(title="Path ")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Path ID", "", "", "", "Average Iterations", "Minimum Iterations", "Maximum Iterations"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"Path  {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # 3: 
    ws3 = wb.create_sheet(title="")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Path ", "", "(x,y,z)", "", "Iterations", "Path "]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"{run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # 4: target paths
    ws4 = wb.create_sheet(title="target paths")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Path ID", "target paths", ""]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"Path  {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f" : {filename}")
    print(f"{'=' * 70}")
    print(f":")
    print(f"  1.        - {len(all_results)} run")
    print(f"  2. Path        - Path ")
    print(f"  3.    -  runPath ")
    print(f"  4. target paths       - target paths")
    print(f"{'=' * 70}\n")

    return filename


def main():


    target_paths = [
        {1, 2, 4, 8, 15, 20, 35, 41, 46, 71, 73, 76, 77, 78, 82, 83, 84, 87, 97, 100, 109, 110, 116, 118, 121, 122, 123,
         125, 126, 131, 132, 133, 134, 137, 141, 144, 155, 156},
        {1, 2, 4, 8, 15, 20, 35, 46, 71, 73, 76, 77, 82, 83, 84, 87, 90, 91, 109, 110, 116, 118, 121, 122, 123, 125,
         126, 131, 132, 133, 134, 137, 140, 141, 144, 155, 156},
        {1, 2, 4, 8, 15, 20, 35, 46, 71, 73, 76, 77, 82, 83, 84, 87, 92, 97, 100, 109, 110, 116, 118, 121, 122, 123,
         125, 126, 131, 132, 133, 134, 137, 141, 144, 155, 156},
        {1, 2, 4, 8, 15, 20, 35, 46, 71, 73, 76, 77, 78, 82, 83, 84, 87, 90, 97, 100, 109, 110, 114, 115, 117, 119, 120,
         124, 127, 128, 134, 137, 140, 141, 144, 155, 156},
        {2, 8, 9, 24, 28, 35, 46, 71, 73, 76, 77, 78, 84, 87, 88, 90, 91, 100, 109, 110, 115, 117, 118, 128, 134, 135,
         141, 144, 146, 147, 148, 150, 151, 153, 154, 155},
        {24, 26, 27, 28, 29, 30, 37, 49, 52, 53, 55, 56, 57, 58, 61, 62, 63, 67, 70, 73, 74, 100, 110, 111, 112, 116,
         117, 126, 129, 130, 131, 141, 142, 144, 155, 156},
        {24, 26, 28, 29, 37, 49, 50, 52, 53, 55, 56, 57, 58, 61, 62, 63, 67, 70, 73, 74, 100, 110, 111, 116, 117, 126,
         129, 130, 131, 134, 141, 142, 144, 155, 156},
        {1, 2, 4, 8, 15, 20, 35, 41, 46, 71, 73, 76, 77, 78, 82, 83, 84, 87, 90, 97, 100, 109, 110, 114, 115, 117, 119,
         120, 124, 127, 128, 134, 137, 152, 157},
        {25, 31, 33, 34, 35, 36, 46, 49, 52, 53, 55, 57, 58, 61, 62, 63, 67, 70, 73, 74, 100, 110, 111, 116, 117, 126,
         129, 130, 131, 141, 142, 144, 155, 156},
        {1, 2, 7, 8, 9, 10, 24, 28, 35, 46, 54, 60, 71, 73, 81, 90, 91, 100, 109, 110, 117, 118, 128, 134, 141, 144,
         146, 147, 148, 150, 151, 153, 154, 155},
        {14, 16, 17, 19, 22, 25, 35, 45, 46, 60, 61, 65, 69, 76, 77, 82, 84, 87, 90, 93, 97, 110, 111, 112, 116, 118,
         126, 129, 130, 131, 144, 155, 156},
        {14, 16, 17, 19, 37, 39, 40, 41, 42, 43, 44, 47, 60, 61, 65, 69, 76, 77, 82, 84, 87, 90, 97, 116, 118, 126, 129,
         131, 134, 137, 144, 155, 156},
        {14, 16, 17, 19, 37, 39, 40, 41, 42, 43, 44, 47, 60, 61, 65, 69, 76, 77, 82, 84, 87, 94, 97, 116, 118, 126, 129,
         131, 134, 137, 144, 155, 156},
        {24, 26, 27, 28, 29, 30, 32, 37, 49, 52, 53, 55, 56, 57, 58, 61, 62, 63, 67, 70, 73, 74, 96, 97, 100, 110, 112,
         128, 141, 142, 144, 155, 156},
        {1, 3, 4, 5, 10, 13, 15, 20, 21, 22, 25, 35, 46, 76, 77, 97, 99, 100, 110, 111, 112, 116, 118, 126, 129, 130,
         131, 141, 142, 144, 155, 156},
        {14, 16, 17, 19, 25, 35, 38, 45, 46, 60, 61, 65, 69, 76, 77, 82, 84, 87, 90, 93, 97, 110, 111, 116, 118, 126,
         129, 130, 131, 144, 155, 156},
        {14, 16, 17, 18, 19, 22, 23, 25, 35, 46, 60, 65, 69, 76, 77, 82, 84, 87, 90, 93, 97, 100, 101, 104, 107, 110,
         111, 112, 128, 144, 155, 156},
        {1, 2, 6, 7, 8, 11, 12, 14, 24, 26, 27, 28, 29, 30, 32, 37, 97, 99, 100, 110, 111, 112, 116, 117, 126, 129, 130,
         131, 134, 141, 142, 144},
        {1, 4, 5, 13, 15, 20, 35, 46, 60, 61, 65, 69, 76, 77, 82, 84, 87, 94, 97, 100, 101, 104, 106, 109, 110, 128,
         134, 137, 144, 155, 156},
        {25, 29, 35, 38, 39, 46, 60, 61, 64, 65, 66, 68, 69, 73, 74, 87, 90, 93, 100, 110, 111, 116, 117, 118, 126, 129,
         131, 144, 155, 156},
        {1, 2, 8, 9, 10, 13, 20, 35, 46, 71, 73, 76, 77, 78, 84, 87, 88, 90, 91, 97, 98, 100, 109, 110, 136, 138, 141,
         142, 144, 155, 156},
        {1, 2, 8, 20, 35, 46, 71, 73, 81, 89, 90, 91, 100, 109, 110, 114, 115, 117, 119, 120, 124, 128, 134, 137, 140,
         141, 144, 155, 156},
        {1, 2, 8, 20, 35, 46, 71, 73, 76, 77, 78, 90, 91, 97, 100, 109, 110, 114, 115, 117, 119, 120, 124, 128, 134,
         137, 143, 145, 152},
        {2, 8, 9, 24, 35, 46, 71, 73, 76, 77, 78, 84, 87, 88, 90, 91, 100, 109, 110, 114, 115, 117, 118, 128, 138, 143,
         145, 149, 152},
        {1, 2, 4, 8, 13, 15, 20, 35, 46, 71, 73, 76, 77, 78, 82, 83, 84, 87, 92, 94, 102, 105, 107, 108, 128, 134, 137,
         144, 155, 156},
        {1, 2, 4, 8, 13, 15, 20, 35, 46, 71, 73, 76, 77, 82, 83, 84, 87, 92, 94, 95, 97, 100, 109, 110, 128, 134, 137,
         144, 155, 156},
        {1, 4, 5, 13, 15, 20, 35, 46, 60, 61, 65, 69, 76, 77, 82, 84, 87, 94, 102, 103, 105, 107, 108, 128, 134, 137,
         144, 155, 156},
        {3, 5, 6, 9, 11, 12, 14, 35, 46, 48, 51, 52, 53, 56, 58, 64, 66, 67, 70, 72, 76, 79, 84, 85, 87, 90, 93, 108,
         144, 155, 156},
        {1, 3, 4, 5, 6, 11, 12, 14, 24, 30, 31, 33, 35, 36, 46, 54, 59, 60, 71, 73, 76, 79, 84, 108, 134, 141, 142, 144,
         155, 156},
        {3, 5, 6, 9, 11, 12, 14, 35, 46, 60, 61, 65, 69, 71, 73, 76, 77, 78, 84, 85, 87, 90, 93, 108, 113, 128, 144,
         155, 156},
        {1, 3, 4, 5, 6, 11, 12, 14, 24, 30, 35, 36, 46, 54, 59, 60, 71, 73, 80, 81, 108, 113, 134, 141, 142, 144, 155,
         156},
        {1, 2, 4, 7, 8, 10, 13, 15, 20, 35, 46, 71, 73, 81, 86, 92, 94, 97, 98, 100, 109, 110, 134, 137, 144, 155, 156},
        {1, 6, 11, 14, 24, 25, 26, 28, 29, 37, 48, 51, 54, 58, 59, 60, 75, 84, 108, 128, 134, 141, 142, 144, 155, 156},
        {2, 7, 8, 9, 10, 24, 28, 35, 46, 71, 73, 81, 90, 91, 100, 109, 110, 117, 118, 128, 139, 141, 144}
    ]

    print("=" * 70)
    print("baseline PSO")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("Program completed")


if __name__ == "__main__":
    main()