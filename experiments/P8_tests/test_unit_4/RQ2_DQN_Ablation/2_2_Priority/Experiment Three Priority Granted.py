import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# ===  ===
dx_min, dx_max = 2, 100  # dx
dy_min, dy_max = 2, 100  # dy
dz_min, dz_max = 2, 100  # dz


# ===  ===
def normalize_state(state):
    """
    [0,1]
    state: (dx, dy, dz)
    """
    dx, dy, dz = state
    normalized_dx = (dx - dx_min) / (dx_max - dx_min)
    normalized_dy = (dy - dy_min) / (dy_max - dy_min)
    normalized_dz = (dz - dz_min) / (dz_max - dz_min)
    return (normalized_dx, normalized_dy, normalized_dz)


def denormalize_state(normalized_state):
    """
    
    """
    norm_dx, norm_dy, norm_dz = normalized_state
    dx = int(norm_dx * (dx_max - dx_min) + dx_min)
    dy = int(norm_dy * (dy_max - dy_min) + dy_min)
    dz = int(norm_dz * (dz_max - dz_min) + dz_min)
    return (dx, dy, dz)


def is_valid_state(state):
    """"""
    dx, dy, dz = state
    return (dx_min <= dx <= dx_max and
            dy_min <= dy <= dy_max and
            dz_min <= dz <= dz_max)


def clip_state(state):
    """"""
    dx, dy, dz = state
    return (
        max(dx_min, min(dx_max, dx)),
        max(dy_min, min(dy_max, dy)),
        max(dz_min, min(dz_max, dz))
    )


# === Metric ===
class PrioritizedMetricsCollector:
    def __init__(self, experiment_name="Prioritized_DQN"):
        self.experiment_name = experiment_name
        self.start_time = None
        self.end_time = None

        # Metric
        self.total_reward = 0
        self.td_errors = []
        self.final_output_similarities = []
        self.action_improvements = []
        self.total_memory_usage = 0
        self.memory_check_count = 0
        self.step_count = 0

        # Metric
        self.episode_rewards = []
        self.episode_similarities = []
        self.episode_td_errors = []
        self.episode_epsilon_values = []
        self.episode_memory_usage = []

        # Metric
        self.priority_statistics = []
        self.importance_weights = []
        self.high_priority_samples_ratio = []
        self.priority_distribution_stats = []

        # Path Metric
        self.similar_paths_performance = []
        self.isolated_paths_performance = []

        # 
        self.milestone_data = {}

        # 
        self.convergence_window = 20
        self.convergence_threshold = 0.02
        self.convergence_detected_episode = None

        # Metric
        self.sample_efficiency_data = []
        self.performance_milestones = [0.6, 0.7, 0.75, 0.8]

        # 
        self.learning_curve_characteristics = {}
        self.early_vs_late_performance = {}

    def reset(self):
        """Metric, """
        self.start_time = None
        self.end_time = None
        self.total_reward = 0
        self.td_errors = []
        self.final_output_similarities = []
        self.action_improvements = []
        self.total_memory_usage = 0
        self.memory_check_count = 0
        self.step_count = 0
        self.episode_rewards = []
        self.episode_similarities = []
        self.episode_td_errors = []
        self.episode_epsilon_values = []
        self.episode_memory_usage = []
        self.priority_statistics = []
        self.importance_weights = []
        self.high_priority_samples_ratio = []
        self.priority_distribution_stats = []
        self.similar_paths_performance = []
        self.isolated_paths_performance = []
        self.milestone_data = {}
        self.convergence_detected_episode = None
        self.sample_efficiency_data = []
        self.learning_curve_characteristics = {}
        self.early_vs_late_performance = {}

    def start_training(self):
        self.start_time = time.time()

    def end_training(self):
        self.end_time = time.time()

    def record_step_metrics(self, reward, td_error, triggered, target_path, priority=None, is_weight=None):
        """Metric, """
        self.step_count += 1
        self.total_reward += reward
        self.td_errors.append(td_error)

        if priority is not None:
            self.priority_statistics.append(priority)
        if is_weight is not None:
            self.importance_weights.append(is_weight)

        process = psutil.Process(os.getpid())
        current_memory = process.memory_info().rss / 1024 / 1024
        self.total_memory_usage += current_memory
        self.memory_check_count += 1

    def record_episode_metrics(self, episode, episode_reward, avg_similarity, avg_td_error, epsilon,
                               path_group="similar", priority_stats=None):
        """episodeMetric, """
        self.episode_rewards.append(episode_reward)
        self.episode_similarities.append(avg_similarity)
        self.episode_td_errors.append(avg_td_error)
        self.episode_epsilon_values.append(epsilon)

        process = psutil.Process(os.getpid())
        current_memory = process.memory_info().rss / 1024 / 1024
        self.episode_memory_usage.append(current_memory)

        if priority_stats:
            self.priority_distribution_stats.append({
                'episode': episode,
                'mean_priority': priority_stats.get('mean_priority', 0),
                'max_priority': priority_stats.get('max_priority', 0),
                'min_priority': priority_stats.get('min_priority', 0),
                'high_priority_ratio': priority_stats.get('high_priority_ratio', 0)
            })

        if path_group == "similar":
            self.similar_paths_performance.append({
                'episode': episode,
                'reward': episode_reward,
                'similarity': avg_similarity,
                'td_error': avg_td_error
            })
        else:
            self.isolated_paths_performance.append({
                'episode': episode,
                'reward': episode_reward,
                'similarity': avg_similarity,
                'td_error': avg_td_error
            })

        if episode in [50, 100, 150, 200, 250, 300, 400, 450, 500]:
            self.milestone_data[episode] = {
                'avg_reward': np.mean(self.episode_rewards[-10:]) if len(
                    self.episode_rewards) >= 10 else episode_reward,
                'avg_similarity': avg_similarity,
                'avg_td_error': avg_td_error,
                'epsilon': epsilon,
                'memory_usage': current_memory,
                'total_steps': self.step_count,
                'priority_stats': priority_stats
            }

        self._check_convergence(episode)
        self._check_performance_milestones(episode, avg_similarity)

    def _check_convergence(self, episode):
        """"""
        if len(self.episode_similarities) >= self.convergence_window and self.convergence_detected_episode is None:
            recent_similarities = self.episode_similarities[-self.convergence_window:]
            if np.std(recent_similarities) < self.convergence_threshold:
                self.convergence_detected_episode = episode

    def _check_performance_milestones(self, episode, similarity):
        """"""
        for milestone in self.performance_milestones:
            if similarity >= milestone and not any(data[1] == milestone for data in self.sample_efficiency_data):
                self.sample_efficiency_data.append((episode, milestone, self.step_count))

    def record_final_output_sample(self, triggered, target_path):
        """final samplesSimilarity"""
        if len(triggered | target_path) > 0:
            similarity = len(triggered & target_path) / len(triggered | target_path)
        else:
            similarity = 0.0
        self.final_output_similarities.append(similarity)

    def record_action_improvement(self, current_reward, prev_reward):
        """"""
        if prev_reward is not None:
            improvement = current_reward - prev_reward
            self.action_improvements.append(1 if improvement > 0 else 0)


# Metric
prioritized_metrics = PrioritizedMetricsCollector("Prioritized_DQN_Enhanced")


# === reward function ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward


def execute_Tr(a):
    """
    替换原有的编排规则函数
    参数a: 包含3个元素的元组或数组，分别对应path_depth, file_count, access_level
    返回: 触发的规则编号集合
    """
    path_depth, file_count, access_level = float(a[0]), int(a[1]), float(a[2])

    # 使用正确的变量名
    x, y, z = path_depth, file_count, access_level

    triggered = set()
    # 创建一个字典来存储b数组的值，用于跟踪哪些规则被触发
    b = {}
    x_score = max(0, 100 - x * 1.67)

    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[0] = 1
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((z * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[1] = 2
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * x) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[2] = 3
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[3] = 4
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * 43) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[4] = 5
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((74 * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[5] = 6
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (47 + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[6] = 7
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (x + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[7] = 8
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (y + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[8] = 9
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 10) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[9] = 10
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) != 50 and x ** 2 + y ** 2 > z ** 2):
        b[10] = 11
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 29 and x ** 2 + y ** 2 > z ** 2):
        b[11] = 12
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 69 and x ** 2 + y ** 2 > z ** 2):
        b[12] = 13
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2): b[
        13] = 14
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and y ** 2 + y ** 2 > z ** 2):
        b[14] = 15
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 3 + y ** 2 > z ** 2):
        b[15] = 16
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + x ** 2 > z ** 2):
        b[16] = 17
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 3 > z ** 2):
        b[17] = 18
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 != z ** 2):
        b[18] = 19
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2.5):
        b[19] = 20
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 1.4 > z ** 2):
        b[20] = 21
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + 34 ** 2 > z ** 2):
        b[21] = 22
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and 64 ** 2 + y ** 2 > z ** 2):
        b[22] = 23
    if (x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2:
        pattern_type = "311A1"
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[23] = 24
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - z ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[24] = 25
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 1.5) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[25] = 26
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2.3 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[26] = 27
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (45 ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[27] = 28
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - 31 ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[28] = 29
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[29] = 30
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (x + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[30] = 31
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 2) < -30 or (abs(x - y) * z) / 100 > 45): b[31] = 32
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (26 + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[32] = 33
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y - 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[33] = 34
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) != -30 or (abs(x - y) * z) / 100 > 45): b[34] = 35
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -10 or (abs(x - y) * z) / 100 > 45): b[35] = 36
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 and (abs(x - y) * z) / 100 > 45): b[36] = 37
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(z - y) * z) / 100 > 45): b[37] = 38
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - z) * z) / 100 > 45): b[38] = 39
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * y) / 100 > 45): b[39] = 40
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * x) / 100 > 45): b[40] = 41
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * 34) / 100 > 45): b[41] = 42
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - 26) * z) / 100 > 45): b[42] = 43
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(52 - y) * z) / 100 > 45): b[43] = 44
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 76 > 45): b[44] = 45
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 != 45): b[45] = 46
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 55): b[46] = 47
    if (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45:
        pattern_type = "311A2"

    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + x) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35): b[47] = 48
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((y + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35): b[48] = 49
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((24 + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35): b[49] = 50
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + 31) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35): b[50] = 51
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 1.5) ** 2 < z * 20 and (x * x * z) / 1000 < 35): b[51] = 52
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2.4 < z * 20 and (x * x * z) / 1000 < 35): b[52] = 53
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 != z * 20 and (x * x * z) / 1000 < 35): b[53] = 54
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < 25 * 20 and (x * x * z) / 1000 < 35): b[54] = 55
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 14 and (x * x * z) / 1000 < 35): b[55] = 56
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < x * 20 and (x * x * z) / 1000 < 35): b[56] = 57
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < y * 20 and (x * x * z) / 1000 < 35): b[57] = 58
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 1.2 < z * 20 and (x * x * z) / 1000 < 35): b[58] = 59
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 or (x * x * z) / 1000 < 35): b[59] = 60
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (y * x * z) / 1000 < 35): b[60] = 61
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (z * x * z) / 1000 < 35): b[61] = 62
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * 35 * z) / 1000 < 35): b[62] = 63
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * x * x) / 1000 < 35): b[63] = 64
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * x * y) / 1000 < 35): b[64] = 65
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * x * 26) / 1000 < 35): b[65] = 66
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * 54 * z) / 1000 < 35): b[66] = 67
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1050 < 35): b[67] = 68
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 != 35): b[68] = 69
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 25): b[69] = 70
    if ((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35:
        pattern_type = "311A3"
    if (x < 40 and z < 50) != (x != 40 and z < 50): b[70] = 71
    if (x < 40 and z < 50) != (x < 29 and z < 50): b[71] = 72
    if (x < 40 and z < 50) != (x < 40 or z < 50): b[72] = 73
    if (x < 40 and z < 50) != (x < 40 and z != 50): b[73] = 74
    if (x < 40 and z < 50) != (x < 40 and z < 36): b[74] = 75
    if x < 40 and z < 50:
        pattern_type = 322
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 or x > 50) or (z < 40 and x > 55 and y < 40)): b[75] = 76
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z != 35 and x > 50) or (z < 40 and x > 55 and y < 40)): b[76] = 77
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 45 and x > 50) or (z < 40 and x > 55 and y < 40)): b[77] = 78
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x != 50) or (z < 40 and x > 55 and y < 40)): b[78] = 79
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 59) or (z < 40 and x > 55 and y < 40)): b[79] = 80
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) and (z < 40 and x > 55 and y < 40)): b[80] = 81
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) or (z != 40 and x > 55 and y < 40)): b[81] = 82
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) or (z < 48 and x > 55 and y < 40)): b[82] = 83
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) or (z < 40 or x > 55 and y < 40)): b[83] = 84
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) or (z < 40 and x != 55 and y < 40)): b[84] = 85
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) or (z < 40 and x > 65 and y < 40)): b[85] = 86
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) or (z < 40 and x > 55 or y < 40)): b[86] = 87
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) or (z < 40 and x > 55 and y != 40)): b[87] = 88
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) or (z < 40 and x > 55 and y < 29)): b[88] = 89
    if (z < 35 and x > 50) or (z < 40 and x > 55 and y < 40):
        pattern_type = 323
    if (y < 25 and x > 60) != (y < 25 or x > 60): b[89] = 90
    if (y < 25 and x > 60) != (y != 25 and x > 60): b[90] = 91
    if (y < 25 and x > 60) != (y < 17 and x > 60): b[91] = 92
    if (y < 25 and x > 60) != (y < 25 and x != 60): b[92] = 93
    if (y < 25 and x > 60) != (y < 25 and x > 69): b[93] = 94
    if y < 25 and x > 60:
        pattern_type = 324
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 75 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)): b[94] = 95
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (22 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)): b[95] = 96
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 or 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)): b[96] = 97
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 29 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)): b[97] = 98
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 78 and (x * z) / 100 < 35) or (x < 60 and z < 60)): b[98] = 99
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 or (x * z) / 100 < 35) or (x < 60 and z < 60)): b[99] = 100
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (56 * z) / 100 < 35) or (x < 60 and z < 60)): b[100] = 101
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * 74) / 100 < 35) or (x < 60 and z < 60)): b[101] = 102
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 89 < 35) or (x < 60 and z < 60)): b[102] = 103
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 != 35) or (x < 60 and z < 60)): b[103] = 104
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 21) or (x < 60 and z < 60)): b[104] = 105
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (z * z) / 100 < 35) or (x < 60 and z < 60)): b[105] = 106
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * x) / 100 < 35) or (x < 60 and z < 60)): b[106] = 107
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) and (x < 60 and z < 60)): b[107] = 108
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x != 60 and z < 60)): b[108] = 109
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 or z < 60)): b[109] = 110
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z != 60)): b[110] = 111
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 68)): b[111] = 112
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 51 and z < 60)): b[112] = 113
    if (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60):
        pattern_type = 325
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((53 ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)): b[113] = 114
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 2.3) / (x + y + 1) < 20) or (z < 65 and x < 75)): b[114] = 115
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 1.4) / (x + y + 1) < 20) or (z < 65 and x < 75)): b[115] = 116
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((x ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)):
        b[116] = 117
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((y ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)):
        b[117] = 118
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((z ** 2) / (y + y + 1) < 20) or (z < 65 and x < 75)):
        b[118] = 119
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((z ** 2) / (z + y + 1) < 20) or (z < 65 and x < 75)):
        b[119] = 120
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((z ** 2) / (x + x + 1) < 20) or (z < 65 and x < 75)):
        b[120] = 121
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((z ** 2) / (x + z + 1) < 20) or (z < 65 and x < 75)):
        b[121] = 122
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 2) / (x + 45 + 1) < 20) or (z < 65 and x < 75)): b[122] = 123
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 2) / (33 + y + 1) < 20) or (z < 65 and x < 75)): b[123] = 124
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 2) / (x + y + 11) < 20) or (z < 65 and x < 75)): b[124] = 125
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 2) / (x + y + 1) != 20) or (z < 65 and x < 75)): b[125] = 126
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((z ** 2) / (x + y + 1) < 15) or (z < 65 and x < 75)):
        b[126] = 127
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 2) / (x + y + 1) < 20) and (z < 65 and x < 75)): b[127] = 128
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 2) / (x + y + 1) < 20) or (z != 65 and x < 75)): b[128] = 129
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((z ** 2) / (x + y + 1) < 20) or (z < 75 and x < 75)):
        b[129] = 130
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((z ** 2) / (x + y + 1) < 20) or (z < 65 or x < 75)):
        b[130] = 131
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 2) / (x + y + 1) < 20) or (z < 65 and x != 75)): b[131] = 132
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 79)):
        b[132] = 133
    if ((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75):
        pattern_type = 326
    if (60 <= x < 82 and y > 50) != (60 <= x < 82 or y > 50): b[133] = 134
    if (60 <= x < 82 and y > 50) != (60 <= x < 89 and y > 50): b[134] = 135
    if (60 <= x < 82 and y > 50) != (67 <= x < 82 and y > 50): b[135] = 136
    if (60 <= x < 82 and y > 50) != (60 <= x < 82 and y != 50): b[136] = 137
    if (60 <= x < 82 and y > 50) != (60 <= x < 82 and y > 62): b[137] = 138
    if 60 <= x < 82 and y > 50:
        pattern_type = 327
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 77 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)): b[138] = 139
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (22 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)): b[139] = 140
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 or x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)): b[140] = 141
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x != 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)): b[141] = 142
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 85 and (y * x) / 100 < 52) or (y < 60 and x > 80)): b[142] = 143
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 or (y * x) / 100 < 52) or (y < 60 and x > 80)): b[143] = 144
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (x * x) / 100 < 52) or (y < 60 and x > 80)): b[144] = 145
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * y) / 100 < 52) or (y < 60 and x > 80)): b[145] = 146
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (43 * x) / 100 < 52) or (y < 60 and x > 80)): b[146] = 147
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * 31) / 100 < 52) or (y < 60 and x > 80)): b[147] = 148
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 78 < 52) or (y < 60 and x > 80)): b[148] = 149
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 100 != 52) or (y < 60 and x > 80)): b[149] = 150
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 100 < 62) or (y < 60 and x > 80)): b[150] = 151
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 100 < 52) and (y < 60 and x > 80)): b[151] = 152
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y != 60 and x > 80)): b[152] = 153
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 65 and x > 80)): b[153] = 154
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 or x > 80)): b[154] = 155
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x != 80)): b[155] = 156
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 88)): b[156] = 157
    if (40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80):
        pattern_type = 328

    # 返回被触发的规则编号集合
    return set(b.values())


def jaccard_similarity(set1, set2):
    """Compute Jaccard similarity"""
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0


# === Path Similarity ===
def compute_path_similarity_matrix(paths):
    """Path Similarity"""
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix


targetPaths = [
    {1, 2, 4, 8, 15, 20, 35, 41, 46, 71, 73, 76, 77, 78, 82, 83, 84, 87, 97, 100, 109, 110, 116, 118, 121, 122, 123,
     125, 126, 131, 132, 133, 134, 137, 141, 144, 155, 156},
    {1, 2, 4, 8, 15, 20, 35, 46, 71, 73, 76, 77, 82, 83, 84, 87, 90, 91, 109, 110, 116, 118, 121, 122, 123, 125, 126,
     131, 132, 133, 134, 137, 140, 141, 144, 155, 156},
    {1, 2, 4, 8, 15, 20, 35, 46, 71, 73, 76, 77, 82, 83, 84, 87, 92, 97, 100, 109, 110, 116, 118, 121, 122, 123, 125,
     126, 131, 132, 133, 134, 137, 141, 144, 155, 156},
    {1, 2, 4, 8, 15, 20, 35, 46, 71, 73, 76, 77, 78, 82, 83, 84, 87, 90, 97, 100, 109, 110, 114, 115, 117, 119, 120,
     124, 127, 128, 134, 137, 140, 141, 144, 155, 156},
    {2, 8, 9, 24, 28, 35, 46, 71, 73, 76, 77, 78, 84, 87, 88, 90, 91, 100, 109, 110, 115, 117, 118, 128, 134, 135, 141,
     144, 146, 147, 148, 150, 151, 153, 154, 155},
    {24, 26, 27, 28, 29, 30, 37, 49, 52, 53, 55, 56, 57, 58, 61, 62, 63, 67, 70, 73, 74, 100, 110, 111, 112, 116, 117,
     126, 129, 130, 131, 141, 142, 144, 155, 156},
    {24, 26, 28, 29, 37, 49, 50, 52, 53, 55, 56, 57, 58, 61, 62, 63, 67, 70, 73, 74, 100, 110, 111, 116, 117, 126, 129,
     130, 131, 134, 141, 142, 144, 155, 156},
    {1, 2, 4, 8, 15, 20, 35, 41, 46, 71, 73, 76, 77, 78, 82, 83, 84, 87, 90, 97, 100, 109, 110, 114, 115, 117, 119, 120,
     124, 127, 128, 134, 137, 152, 157},
    {25, 31, 33, 34, 35, 36, 46, 49, 52, 53, 55, 57, 58, 61, 62, 63, 67, 70, 73, 74, 100, 110, 111, 116, 117, 126, 129,
     130, 131, 141, 142, 144, 155, 156},
    {1, 2, 7, 8, 9, 10, 24, 28, 35, 46, 54, 60, 71, 73, 81, 90, 91, 100, 109, 110, 117, 118, 128, 134, 141, 144, 146,
     147, 148, 150, 151, 153, 154, 155},
    {14, 16, 17, 19, 22, 25, 35, 45, 46, 60, 61, 65, 69, 76, 77, 82, 84, 87, 90, 93, 97, 110, 111, 112, 116, 118, 126,
     129, 130, 131, 144, 155, 156},
    {14, 16, 17, 19, 37, 39, 40, 41, 42, 43, 44, 47, 60, 61, 65, 69, 76, 77, 82, 84, 87, 90, 97, 116, 118, 126, 129,
     131, 134, 137, 144, 155, 156},
    {14, 16, 17, 19, 37, 39, 40, 41, 42, 43, 44, 47, 60, 61, 65, 69, 76, 77, 82, 84, 87, 94, 97, 116, 118, 126, 129,
     131, 134, 137, 144, 155, 156},
    {24, 26, 27, 28, 29, 30, 32, 37, 49, 52, 53, 55, 56, 57, 58, 61, 62, 63, 67, 70, 73, 74, 96, 97, 100, 110, 112, 128,
     141, 142, 144, 155, 156},
    {1, 3, 4, 5, 10, 13, 15, 20, 21, 22, 25, 35, 46, 76, 77, 97, 99, 100, 110, 111, 112, 116, 118, 126, 129, 130, 131,
     141, 142, 144, 155, 156},
    {14, 16, 17, 19, 25, 35, 38, 45, 46, 60, 61, 65, 69, 76, 77, 82, 84, 87, 90, 93, 97, 110, 111, 116, 118, 126, 129,
     130, 131, 144, 155, 156},
    {14, 16, 17, 18, 19, 22, 23, 25, 35, 46, 60, 65, 69, 76, 77, 82, 84, 87, 90, 93, 97, 100, 101, 104, 107, 110, 111,
     112, 128, 144, 155, 156},
    {1, 2, 6, 7, 8, 11, 12, 14, 24, 26, 27, 28, 29, 30, 32, 37, 97, 99, 100, 110, 111, 112, 116, 117, 126, 129, 130,
     131, 134, 141, 142, 144},
    {1, 4, 5, 13, 15, 20, 35, 46, 60, 61, 65, 69, 76, 77, 82, 84, 87, 94, 97, 100, 101, 104, 106, 109, 110, 128, 134,
     137, 144, 155, 156},
    {25, 29, 35, 38, 39, 46, 60, 61, 64, 65, 66, 68, 69, 73, 74, 87, 90, 93, 100, 110, 111, 116, 117, 118, 126, 129,
     131, 144, 155, 156},
    {1, 2, 8, 9, 10, 13, 20, 35, 46, 71, 73, 76, 77, 78, 84, 87, 88, 90, 91, 97, 98, 100, 109, 110, 136, 138, 141, 142,
     144, 155, 156},
    {1, 2, 8, 20, 35, 46, 71, 73, 81, 89, 90, 91, 100, 109, 110, 114, 115, 117, 119, 120, 124, 128, 134, 137, 140, 141,
     144, 155, 156},
    {1, 2, 8, 20, 35, 46, 71, 73, 76, 77, 78, 90, 91, 97, 100, 109, 110, 114, 115, 117, 119, 120, 124, 128, 134, 137,
     143, 145, 152},
    {2, 8, 9, 24, 35, 46, 71, 73, 76, 77, 78, 84, 87, 88, 90, 91, 100, 109, 110, 114, 115, 117, 118, 128, 138, 143, 145,
     149, 152},
    {1, 2, 4, 8, 13, 15, 20, 35, 46, 71, 73, 76, 77, 78, 82, 83, 84, 87, 92, 94, 102, 105, 107, 108, 128, 134, 137, 144,
     155, 156},
    {1, 2, 4, 8, 13, 15, 20, 35, 46, 71, 73, 76, 77, 82, 83, 84, 87, 92, 94, 95, 97, 100, 109, 110, 128, 134, 137, 144,
     155, 156},
    {1, 4, 5, 13, 15, 20, 35, 46, 60, 61, 65, 69, 76, 77, 82, 84, 87, 94, 102, 103, 105, 107, 108, 128, 134, 137, 144,
     155, 156},
    {3, 5, 6, 9, 11, 12, 14, 35, 46, 48, 51, 52, 53, 56, 58, 64, 66, 67, 70, 72, 76, 79, 84, 85, 87, 90, 93, 108, 144,
     155, 156},
    {1, 3, 4, 5, 6, 11, 12, 14, 24, 30, 31, 33, 35, 36, 46, 54, 59, 60, 71, 73, 76, 79, 84, 108, 134, 141, 142, 144,
     155, 156},
    {3, 5, 6, 9, 11, 12, 14, 35, 46, 60, 61, 65, 69, 71, 73, 76, 77, 78, 84, 85, 87, 90, 93, 108, 113, 128, 144, 155,
     156},
    {1, 3, 4, 5, 6, 11, 12, 14, 24, 30, 35, 36, 46, 54, 59, 60, 71, 73, 80, 81, 108, 113, 134, 141, 142, 144, 155, 156},
    {1, 2, 4, 7, 8, 10, 13, 15, 20, 35, 46, 71, 73, 81, 86, 92, 94, 97, 98, 100, 109, 110, 134, 137, 144, 155, 156},
    {1, 6, 11, 14, 24, 25, 26, 28, 29, 37, 48, 51, 54, 58, 59, 60, 75, 84, 108, 128, 134, 141, 142, 144, 155, 156},
    {2, 7, 8, 9, 10, 24, 28, 35, 46, 71, 73, 81, 90, 91, 100, 109, 110, 117, 118, 128, 139, 141, 144}
]


# === Path  ===
def group_paths_by_similarity(paths):
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.mean(avg_sim_scores)

    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]
    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)

    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group


# === ()===
class PrioritizedExperienceReplay:
    def __init__(self, capacity=10000, alpha=0.6, beta_start=0.4, beta_frames=100000):
        self.capacity = capacity
        self.alpha = alpha
        self.beta_start = beta_start
        self.beta_frames = beta_frames
        self.frame = 1

        self.buffer = []
        self.priorities = np.zeros(capacity, dtype=np.float32)
        self.pos = 0
        self.size = 0

        self.max_priority = 1.0
        self.min_priority = 1.0

    def beta(self):
        """beta()"""
        return min(1.0, self.beta_start + (1.0 - self.beta_start) * self.frame / self.beta_frames)

    def append(self, experience):
        """, """
        if len(self.buffer) < self.capacity:
            self.buffer.append(experience)
        else:
            self.buffer[self.pos] = experience

        self.priorities[self.pos] = self.max_priority

        self.pos = (self.pos + 1) % self.capacity
        self.size = min(self.size + 1, self.capacity)

    def sample(self, batch_size):
        """, """
        if self.size < batch_size:
            return [], [], []

        priorities = self.priorities[:self.size]
        probs = priorities ** self.alpha
        probs /= probs.sum()

        # replace=False
        indices = np.random.choice(self.size, batch_size, p=probs, replace=False)

        # : 
        unique_batch = []
        unique_indices = []
        seen_states = set()

        for idx in indices:
            experience = self.buffer[idx]
            state_tensor = experience[0]
            # 
            state_tuple = tuple(state_tensor.cpu().numpy().flatten())

            if state_tuple not in seen_states:
                seen_states.add(state_tuple)
                unique_batch.append(experience)
                unique_indices.append(idx)

        # , 
        if len(unique_batch) < batch_size:
            remaining_indices = [i for i in range(self.size) if i not in unique_indices]
            if remaining_indices:
                remaining_probs = priorities[remaining_indices] ** self.alpha
                remaining_probs /= remaining_probs.sum()

                needed = batch_size - len(unique_batch)
                additional_indices = np.random.choice(
                    remaining_indices,
                    min(needed, len(remaining_indices)),
                    p=remaining_probs,
                    replace=False
                )

                for idx in additional_indices:
                    experience = self.buffer[idx]
                    state_tensor = experience[0]
                    state_tuple = tuple(state_tensor.cpu().numpy().flatten())

                    if state_tuple not in seen_states:
                        seen_states.add(state_tuple)
                        unique_batch.append(experience)
                        unique_indices.append(idx)

        total = len(self.buffer)
        unique_indices = np.array(unique_indices)
        weights = (total * probs[unique_indices]) ** (-self.beta())
        weights /= weights.max()

        self.frame += 1

        return unique_batch, unique_indices, weights

    def update_priorities(self, indices, priorities):
        """"""
        for idx, priority in zip(indices, priorities):
            if idx < self.size:
                self.priorities[idx] = priority
                self.max_priority = max(self.max_priority, priority)
                self.min_priority = min(self.min_priority, priority)

    def get_priority_statistics(self):
        """"""
        if self.size == 0:
            return None

        priorities = self.priorities[:self.size]
        mean_priority = np.mean(priorities)
        max_priority = np.max(priorities)
        min_priority = np.min(priorities)

        high_priority_ratio = np.mean(priorities > mean_priority)

        return {
            'mean_priority': mean_priority,
            'max_priority': max_priority,
            'min_priority': min_priority,
            'high_priority_ratio': high_priority_ratio
        }

    def __len__(self):
        return self.size

    def get_high_reward_samples(self, target_path, num_samples=20):
        """"""
        if len(self.buffer) == 0:
            return []

        samples_with_recalculated_scores = []
        for experience in self.buffer:
            state_tensor = experience[0]
            state_tuple = tuple(state_tensor.cpu().numpy().flatten().astype(int))
            triggered = execute_Tr(*state_tuple)  #  dx, dy, dz
            new_reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_recalculated_scores.append((state_tuple, new_reward, sim, triggered))

        samples_with_recalculated_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_recalculated_scores[:num_samples]


def load_path_data(file_path):
    path_data = []
    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines[2:]:
            parts = line.strip().split("\t")
            state = tuple(map(int, parts[0].split()))
            path_data.append(state)
    return path_data


# === DQN ===
class DQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, 128)
        self.fc2 = nn.Linear(128, 64)
        self.fc3 = nn.Linear(64, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)


# === DQN Agent()===
class PrioritizedDQNAgent:
    def __init__(self, state_dim, action_dim, replay_buffer, gamma=0.99, epsilon=1.0,
                 epsilon_decay=0.995, epsilon_min=0.1, learning_rate=0.001):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        """
        
        : 30 = 3 x 10
        - 0: dx (1-50)
        - 1: dy (1-50)
        - 2: dz (1-50)

        : 
        - dx, dy, dz: +/-35(70%), +/-25(50%), +/-10(20%), +/-5(10%), +/-2(5%)(50)
        """
        # : 50
        delta_values = [35, 25, 10, 5, 2, -2, -5, -10, -25, -35]

        dim = action_idx // 10
        delta_idx = action_idx % 10

        if dim == 0:  # dx
            return (delta_values[delta_idx], 0, 0)
        elif dim == 1:  # dy
            return (0, delta_values[delta_idx], 0)
        elif dim == 2:  # dz
            return (0, 0, delta_values[delta_idx])
        else:
            return (0, 0, 0)

    def act(self, state):
        if random.random() < self.epsilon:
            return random.randrange(self.action_dim)
        # act
        normalized_state = normalize_state(state)
        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)
        return torch.argmax(q_values, dim=1).item()

    def store_transition(self, state, action, reward, next_state, done):
        """, TD()"""
        # 
        normalized_state = normalize_state(state)
        normalized_next_state = normalize_state(next_state)

        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        next_state_tensor = torch.tensor(normalized_next_state, dtype=torch.float32).unsqueeze(0).to(device)

        with torch.no_grad():
            q_values = self.model(state_tensor)
            next_q_values = self.target_model(next_state_tensor)
            max_next_q_values = next_q_values.max(1)[0]
            target_q_values = reward + (self.gamma * max_next_q_values * (1 - done))
            td_error = torch.abs(q_values[0][action] - target_q_values).item()

        self.replay_buffer.append((state_tensor, action, reward, next_state_tensor, done, td_error))
        return td_error

    def train(self, batch_size=32):
        """()"""
        if len(self.replay_buffer) < batch_size:
            return

        batch, batch_indices, is_weights = self.replay_buffer.sample(batch_size)
        if not batch:
            return

        states, actions, rewards, next_states, dones, _ = zip(*batch)

        # tensor, 
        states = torch.tensor(np.array([s.cpu().numpy().flatten() for s in states]), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array([ns.cpu().numpy().flatten() for ns in next_states]),
                                   dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)
        is_weights = torch.tensor(is_weights, dtype=torch.float32).to(device)

        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))

        td_errors = torch.abs(current_q_values - target_q_values)
        loss = (is_weights * (current_q_values - target_q_values) ** 2).mean()

        self.optimizer.zero_grad()
        loss.backward()
        self.optimizer.step()

        new_priorities = td_errors.detach().cpu().numpy() + 1e-6
        self.replay_buffer.update_priorities(batch_indices, new_priorities)

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())


# === Sample generation===
def generate_samples_for_similar_paths(similar_group_indices, num_total=2000, top_k=200):
    def jaccard_similarity_local(a, b):
        if not a and not b:
            return 1.0
        return len(a & b) / len(a | b) if a | b else 0.0

    def compute_robustness(state, path):
        base = execute_Tr(*state)  #  dx, dy, dz
        if not base:
            return 0.0
        rob, neighbors = 0.0, 0
        for ddx in [-1, 0, 1]:
            for ddy in [-1, 0, 1]:
                for ddz in [-1, 0, 1]:
                    if ddx == ddy == ddz == 0:
                        continue
                    neighbor_state = (state[0] + ddx, state[1] + ddy, state[2] + ddz)
                    if not is_valid_state(neighbor_state):
                        continue
                    neighbor = clip_state(neighbor_state)
                    n_trig = execute_Tr(*neighbor)  #  dx, dy, dz
                    if not n_trig:
                        continue
                    rob += jaccard_similarity_local(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"prioritized_path{path_id}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Prioritized Path {path_id}\n")
            f.write("dx dy dz\tScore\tSimilarity\tLengthDiff\tRobustness\n")
            for s in samples:
                dx, dy, dz = s[0]
                f.write(f"{dx} {dy} {dz}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_prioritized"
    for path_idx in similar_group_indices:
        path = targetPaths[path_idx]
        samples = []
        attempts = 0
        while len(samples) < top_k and attempts < num_total * 5:
            attempts += 1
            state = (
                np.random.randint(dx_min, dx_max + 1),
                np.random.randint(dy_min, dy_max + 1),
                np.random.randint(dz_min, dz_max + 1)
            )
            triggered = execute_Tr(*state)  #  dx, dy, dz
            if not triggered:
                continue
            sim = jaccard_similarity_local(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            score = 0.55 * sim + 0.25 * len_diff + 0.2 * rob
            samples.append((state, score, sim, len_diff, rob))
        if samples:
            samples.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples[:top_k], base_dir=base_dir)


# === Run ===
def prioritized_generate_and_train_for_similar_paths(agent, similar_group, path_documents, episodes=500,
                                                     batch_size=32, steps_per_test=3, replay_times=1,
                                                     is_isolated=False):
    trained_paths = set()
    global_replay_count = 0

    for episode in range(episodes):
        episode_reward = 0
        episode_similarities = []
        episode_td_errors = []

        for path_idx in similar_group:
            if path_idx in trained_paths:
                continue

            file_path = os.path.join(path_documents,
                                     f"prioritized_path{path_idx + 1}{'_isolated' if is_isolated else ''}.txt")
            path_data = load_path_data(file_path)
            target_path = targetPaths[path_idx]

            # 
            BATCH_SIZE = 50
            N_SAMPLES = 200
            N_STEPS = 3
            N_BATCHES = 4
            PATH_REPEAT = 5

            for repeat in range(PATH_REPEAT):
                for batch_idx in range(N_BATCHES):
                    batch_start = batch_idx * BATCH_SIZE
                    batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)

                    for test_data_idx in range(batch_start, batch_end):
                        if test_data_idx >= len(path_data):
                            break

                        state = path_data[test_data_idx]
                        prev_state = None
                        prev_triggered = None
                        prev_reward = None

                        for step in range(N_STEPS):
                            legal_actions = []
                            for a in range(agent.action_dim):
                                ddx, ddy, ddz = agent.decode_action(a)
                                cand_next = (state[0] + ddx, state[1] + ddy, state[2] + ddz)
                                if is_valid_state(cand_next):
                                    legal_actions.append(a)
                            if not legal_actions:
                                break

                            if random.random() < agent.epsilon:
                                action = random.choice(legal_actions)
                            else:
                                # 
                                normalized_state = normalize_state(state)
                                state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(
                                    device)
                                with torch.no_grad():
                                    q_values = agent.model(state_tensor)[0]
                                action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                            ddx, ddy, ddz = agent.decode_action(action)
                            next_state = clip_state((state[0] + ddx, state[1] + ddy, state[2] + ddz))

                            triggered = execute_Tr(*next_state)  #  dx, dy, dz
                            reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                            done = (step == N_STEPS - 1)

                            td_error = agent.store_transition(state, action, reward, next_state, done)

                            priority_stats = agent.replay_buffer.get_priority_statistics()
                            current_priority = priority_stats['mean_priority'] if priority_stats else 0

                            prioritized_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                    current_priority, None)

                            episode_similarities.append(jaccard_similarity(triggered, target_path))
                            episode_td_errors.append(td_error)

                            if prev_reward is not None:
                                prioritized_metrics.record_action_improvement(reward, prev_reward)

                            prev_state = state
                            prev_triggered = triggered
                            prev_reward = reward
                            state = next_state
                            episode_reward += reward

                    if len(agent.replay_buffer) >= batch_size:
                        agent.train(batch_size)
                        global_replay_count += 1

                        if global_replay_count % 2 == 0:
                            agent.update_target_model()

            trained_paths.add(path_idx)

        avg_similarity = np.mean(episode_similarities) if episode_similarities else 0
        avg_td_error = np.mean(episode_td_errors) if episode_td_errors else 0
        priority_stats = agent.replay_buffer.get_priority_statistics()

        prioritized_metrics.record_episode_metrics(episode, episode_reward, avg_similarity, avg_td_error,
                                                   agent.epsilon, "similar", priority_stats)

        if len(trained_paths) == len(similar_group):
            break

    return agent


def generate_samples_for_isolated_paths_prioritized(agent_similar, isolated_group_indices, num_total=2000, top_k=200):
    def compute_q_value_normalized_complement(state, agent):
        """Q"""
        # 
        normalized_state = normalize_state(state)
        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = agent.model(state_tensor)

        all_q_values = q_values[0].cpu().numpy()

        q_min = all_q_values.min()
        q_max = all_q_values.max()

        if q_max - q_min > 1e-6:
            normalized_q = (all_q_values.max() - q_min) / (q_max - q_min)
        else:
            normalized_q = 0.0

        complement_q = 1.0 - normalized_q

        return complement_q

    def compute_robustness(state, path):
        base = execute_Tr(*state)  #  dx, dy, dz
        if not base:
            return 0.0
        rob, neighbors = 0.0, 0
        for ddx in [-1, 0, 1]:
            for ddy in [-1, 0, 1]:
                for ddz in [-1, 0, 1]:
                    if ddx == ddy == ddz == 0:
                        continue
                    neighbor_state = (state[0] + ddx, state[1] + ddy, state[2] + ddz)
                    if not is_valid_state(neighbor_state):
                        continue
                    neighbor = clip_state(neighbor_state)
                    n_trig = execute_Tr(*neighbor)  #  dx, dy, dz
                    if not n_trig:
                        continue
                    rob += jaccard_similarity(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"prioritized_path{path_id}_isolated.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Prioritized Isolated Path {path_id}\n")
            f.write("dx dy dz\tScore\tSimilarity\tLengthDiff\tRobustness\tQ_value_complement\n")
            for s in samples:
                dx, dy, dz = s[0]
                f.write(f"{dx} {dy} {dz}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\t{s[5]:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_prioritized"
    for path_idx in isolated_group_indices:
        path = targetPaths[path_idx]
        samples = []
        attempts = 0
        while len(samples) < top_k and attempts < num_total * 5:
            attempts += 1
            state = (
                np.random.randint(dx_min, dx_max + 1),
                np.random.randint(dy_min, dy_max + 1),
                np.random.randint(dz_min, dz_max + 1)
            )
            triggered = execute_Tr(*state)  #  dx, dy, dz
            if not triggered:
                continue
            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            q_value_complement = compute_q_value_normalized_complement(state, agent_similar)
            score = 0.28 * sim + 0.1 * len_diff + 0.19 * rob + 0.43 * q_value_complement
            samples.append((state, score, sim, len_diff, rob, q_value_complement))
        if samples:
            samples.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples[:top_k], base_dir=base_dir)


# === Run ===
def prioritized_generate_and_train_for_isolated_paths(agent_similar, agent_isolated, similar_group,
                                                      isolated_group, path_documents, episodes=500, batch_size=32,
                                                      is_isolated=True):
    trained_paths = set()
    global_replay_count = 0

    stage1_samples_pool = {}
    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_similar.replay_buffer.get_high_reward_samples(target_path, num_samples=100)
        stage1_samples_pool[path_idx] = high_reward_samples

    for episode in range(episodes):
        episode_reward = 0
        episode_similarities = []
        episode_td_errors = []

        for path_idx in isolated_group:
            if path_idx in trained_paths:
                continue

            file_path = os.path.join(path_documents, f"prioritized_path{path_idx + 1}_isolated.txt")
            stage2_path_data = load_path_data(file_path)
            stage1_samples = stage1_samples_pool.get(path_idx, [])
            target_path = targetPaths[path_idx]

            # 
            BATCH_SIZE = 50
            N_SAMPLES_STAGE2 = min(140, len(stage2_path_data))
            N_SAMPLES_STAGE1 = min(60, len(stage1_samples))
            N_STEPS = 3
            N_BATCHES_STAGE2 = (N_SAMPLES_STAGE2 + BATCH_SIZE - 1) // BATCH_SIZE
            N_BATCHES_STAGE1 = (N_SAMPLES_STAGE1 + BATCH_SIZE - 1) // BATCH_SIZE
            PATH_REPEAT = 5

            for repeat in range(PATH_REPEAT):
                # Stage2
                for batch_idx in range(N_BATCHES_STAGE2):
                    batch_start = batch_idx * BATCH_SIZE
                    batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES_STAGE2)

                    for test_data_idx in range(batch_start, batch_end):
                        if test_data_idx >= len(stage2_path_data):
                            break

                        state = stage2_path_data[test_data_idx]
                        prev_state = None
                        prev_triggered = None
                        prev_reward = None

                        for step in range(N_STEPS):
                            legal_actions = []
                            for a in range(agent_isolated.action_dim):
                                ddx, ddy, ddz = agent_isolated.decode_action(a)
                                cand_next = (state[0] + ddx, state[1] + ddy, state[2] + ddz)
                                if is_valid_state(cand_next):
                                    legal_actions.append(a)

                            if not legal_actions:
                                break

                            if random.random() < agent_isolated.epsilon:
                                action = random.choice(legal_actions)
                            else:
                                # 
                                normalized_state = normalize_state(state)
                                state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(
                                    device)
                                with torch.no_grad():
                                    q_values = agent_isolated.model(state_tensor)[0]
                                action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                            ddx, ddy, ddz = agent_isolated.decode_action(action)
                            next_state = clip_state((state[0] + ddx, state[1] + ddy, state[2] + ddz))

                            triggered = execute_Tr(*next_state)  #  dx, dy, dz
                            reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)

                            if target_path.issubset(triggered):
                                reward += 2.0

                            done = (step == N_STEPS - 1)

                            td_error = agent_isolated.store_transition(state, action, reward, next_state, done)

                            priority_stats = agent_isolated.replay_buffer.get_priority_statistics()
                            current_priority = priority_stats['mean_priority'] if priority_stats else 0

                            prioritized_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                    current_priority, None)

                            episode_similarities.append(jaccard_similarity(triggered, target_path))
                            episode_td_errors.append(td_error)

                            if prev_reward is not None:
                                prioritized_metrics.record_action_improvement(reward, prev_reward)

                            prev_state = state
                            prev_triggered = triggered
                            prev_reward = reward
                            state = next_state
                            episode_reward += reward

                    if len(agent_isolated.replay_buffer) >= batch_size:
                        agent_isolated.train(batch_size)
                        global_replay_count += 1

                        if global_replay_count % 2 == 0:
                            agent_isolated.update_target_model()

                # Stage1
                if stage1_samples:
                    for batch_idx in range(N_BATCHES_STAGE1):
                        batch_start = batch_idx * BATCH_SIZE
                        batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES_STAGE1)

                        for sample_idx in range(batch_start, batch_end):
                            if sample_idx >= len(stage1_samples):
                                break

                            stage1_state_tuple, _, _, _ = stage1_samples[sample_idx]
                            state = stage1_state_tuple
                            prev_state = None
                            prev_triggered = None
                            prev_reward = None

                            for step in range(N_STEPS):
                                legal_actions = []
                                for a in range(agent_isolated.action_dim):
                                    ddx, ddy, ddz = agent_isolated.decode_action(a)
                                    cand_next = (state[0] + ddx, state[1] + ddy, state[2] + ddz)
                                    if is_valid_state(cand_next):
                                        legal_actions.append(a)

                                if not legal_actions:
                                    break

                                if random.random() < agent_isolated.epsilon:
                                    action = random.choice(legal_actions)
                                else:
                                    # 
                                    normalized_state = normalize_state(state)
                                    state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(
                                        device)
                                    with torch.no_grad():
                                        q_values = agent_isolated.model(state_tensor)[0]
                                    action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                                ddx, ddy, ddz = agent_isolated.decode_action(action)
                                next_state = clip_state((state[0] + ddx, state[1] + ddy, state[2] + ddz))

                                triggered = execute_Tr(*next_state)  #  dx, dy, dz
                                reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                                reward *= 0.8

                                done = (step == N_STEPS - 1)

                                td_error = agent_isolated.store_transition(state, action, reward, next_state, done)

                                priority_stats = agent_isolated.replay_buffer.get_priority_statistics()
                                current_priority = priority_stats['mean_priority'] if priority_stats else 0

                                prioritized_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                        current_priority, None)

                                episode_similarities.append(jaccard_similarity(triggered, target_path))
                                episode_td_errors.append(td_error)

                                if prev_reward is not None:
                                    prioritized_metrics.record_action_improvement(reward, prev_reward)

                                prev_state = state
                                prev_triggered = triggered
                                prev_reward = reward
                                state = next_state
                                episode_reward += reward

                        if len(agent_isolated.replay_buffer) >= batch_size:
                            agent_isolated.train(batch_size)
                            global_replay_count += 1

                            if global_replay_count % 2 == 0:
                                agent_isolated.update_target_model()

            trained_paths.add(path_idx)

        avg_similarity = np.mean(episode_similarities) if episode_similarities else 0
        avg_td_error = np.mean(episode_td_errors) if episode_td_errors else 0
        priority_stats = agent_isolated.replay_buffer.get_priority_statistics()

        prioritized_metrics.record_episode_metrics(episode, episode_reward, avg_similarity, avg_td_error,
                                                   agent_isolated.epsilon, "isolated", priority_stats)

        if len(trained_paths) == len(isolated_group):
            break

    return agent_isolated


# === Excel()===
def append_performance_metrics_to_excel(metrics_collector, filepath, run_number):
    """MetricExcel"""
    avg_td_error = np.mean(metrics_collector.td_errors) if metrics_collector.td_errors else 0
    total_reward = metrics_collector.total_reward
    action_improvement_rate = np.mean(
        metrics_collector.action_improvements) if metrics_collector.action_improvements else 0
    avg_final_similarity = np.mean(
        metrics_collector.final_output_similarities) if metrics_collector.final_output_similarities else 0
    avg_episode_reward = np.mean(metrics_collector.episode_rewards) if metrics_collector.episode_rewards else 0
    reward_std = np.std(metrics_collector.episode_rewards) if metrics_collector.episode_rewards else 0
    training_time = metrics_collector.end_time - metrics_collector.start_time if metrics_collector.end_time and metrics_collector.start_time else 0
    avg_memory_usage = np.mean(metrics_collector.episode_memory_usage) if metrics_collector.episode_memory_usage else 0
    per_step_time = training_time / metrics_collector.step_count * 1000 if metrics_collector.step_count > 0 else 0
    avg_priority = np.mean(metrics_collector.priority_statistics) if metrics_collector.priority_statistics else 0
    avg_importance_weight = np.mean(metrics_collector.importance_weights) if metrics_collector.importance_weights else 0

    new_row = {
        'Run': f"Run {run_number}",
        'final samplesAverage Similarity': f"{avg_final_similarity:.4f}",
        '': f"{total_reward:,.2f}",
        'episode': f"{avg_episode_reward:,.4f}",
        'Standard deviation': f"{reward_std:,.4f}",
        'TD': f"{avg_td_error:.4f}",
        '': f"{action_improvement_rate:.4f}",
        '(%)': f"{action_improvement_rate * 100:.2f}%",
        '': f"{metrics_collector.step_count:,}",
        'final samples': f"{len(metrics_collector.final_output_similarities)}",
        'Training Total Time( seconds)': f"{training_time:.2f}",
        'Training Total Time( minutes)': f"{training_time / 60:.2f}",
        '(MB)': f"{avg_memory_usage:.2f}",
        '(ms)': f"{per_step_time:.2f}",
        '': f"{avg_priority:.4f}",
        '': f"{avg_importance_weight:.4f}"
    }

    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    if os.path.exists(filepath):
        df = pd.read_excel(filepath, sheet_name='Metric')
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    else:
        df = pd.DataFrame([new_row])

    with pd.ExcelWriter(filepath, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, sheet_name='Metric', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Metric']

        worksheet.column_dimensions['A'].width = 15
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
            worksheet.column_dimensions[col].width = 18

        header_font = Font(bold=True, size=11)
        for cell in worksheet[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    print(f"Run {run_number} runMetric: {filepath}")


def append_final_samples_to_excel(agent_similar, agent_isolated, similar_group, isolated_group, targetPaths, filepath,
                                  run_number):
    """Excel"""
    new_samples = []

    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_similar.replay_buffer.get_high_reward_samples(target_path, num_samples=20)

        for state_tuple, reward, sim, triggered in high_reward_samples:
            new_samples.append({
                'Run': f"Run {run_number}",
                'Path ': 'Similar path group',
                'Path ID': path_idx + 1,
                'dx': state_tuple[0],
                'dy': state_tuple[1],
                'dz': state_tuple[2],
                'Similarity': f"{sim:.4f}",
                '': f"{reward:.2f}",
                '': len(triggered),
                '': len(target_path),
                '': str(sorted(triggered)),
                '': str(sorted(target_path))
            })

    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)

        for state_tuple, reward, sim, triggered in high_reward_samples:
            new_samples.append({
                'Run': f"Run {run_number}",
                'Path ': 'Isolated path group',
                'Path ID': path_idx + 1,
                'dx': state_tuple[0],
                'dy': state_tuple[1],
                'dz': state_tuple[2],
                'Similarity': f"{sim:.4f}",
                '': f"{reward:.2f}",
                '': len(triggered),
                '': len(target_path),
                '': str(sorted(triggered)),
                '': str(sorted(target_path))
            })

    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    if os.path.exists(filepath):
        df = pd.read_excel(filepath, sheet_name='final samples')
        df = pd.concat([df, pd.DataFrame(new_samples)], ignore_index=True)
    else:
        df = pd.DataFrame(new_samples)

    with pd.ExcelWriter(filepath, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, sheet_name='final samples', index=False)
        workbook = writer.book
        worksheet = writer.sheets['final samples']

        column_widths = {
            'A': 12, 'B': 15, 'C': 12, 'D': 10, 'E': 10, 'F': 10,
            'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 40, 'L': 40
        }
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        header_font = Font(bold=True, size=11)
        for cell in worksheet[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    print(f"Run {run_number} run: {filepath}")


# ===  run ===
def run_single_experiment(run_number, results_save_dir):
    """"""
    print(f"\n{'=' * 80}")
    print(f"Start run  {run_number}  run(DQN)")
    print(f"{'=' * 80}\n")

    # Metric
    prioritized_metrics.reset()
    prioritized_metrics.start_training()

    model_path_similar = os.path.join(results_save_dir, f"run{run_number}_similar.pth")
    model_path_isolated = os.path.join(results_save_dir, f"run{run_number}_isolated.pth")
    path_documents = r"D:\Experiment\CNN\DQNNEW\path_samples_prioritized"

    similar_group, isolated_group = group_paths_by_similarity(targetPaths)

    # Run : Path 
    if run_number == 1:
        generate_samples_for_similar_paths(similar_group, num_total=2000, top_k=200)

    replay_buffer = PrioritizedExperienceReplay(capacity=10000, alpha=0.6, beta_start=0.4, beta_frames=100000)
    state_dim = 3
    action_dim = 30
    agent = PrioritizedDQNAgent(state_dim, action_dim, replay_buffer)

    agent = prioritized_generate_and_train_for_similar_paths(agent, similar_group, path_documents, episodes=500,
                                                             batch_size=32, is_isolated=False)

    os.makedirs(os.path.dirname(model_path_similar), exist_ok=True)
    torch.save({
        'model_state_dict': agent.model.state_dict(),
        'optimizer_state_dict': agent.optimizer.state_dict(),
        'epsilon': agent.epsilon
    }, model_path_similar)

    # Run : Path 
    if run_number == 1:
        generate_samples_for_isolated_paths_prioritized(agent, isolated_group, num_total=2000, top_k=200)

    # Run : Path 
    isolated_replay_buffer = PrioritizedExperienceReplay(capacity=15000, alpha=0.6, beta_start=0.4, beta_frames=100000)
    agent_isolated = PrioritizedDQNAgent(state_dim, action_dim, isolated_replay_buffer)

    try:
        checkpoint = torch.load(model_path_similar)
        agent_isolated.model.load_state_dict(checkpoint['model_state_dict'])
        agent_isolated.target_model.load_state_dict(checkpoint['model_state_dict'])
        agent_isolated.epsilon = checkpoint.get('epsilon', 0.5)
    except Exception as e:
        pass

    agent_isolated = prioritized_generate_and_train_for_isolated_paths(
        agent_similar=agent,
        agent_isolated=agent_isolated,
        similar_group=similar_group,
        isolated_group=isolated_group,
        path_documents=path_documents,
        episodes=500,
        batch_size=32,
        is_isolated=True
    )

    os.makedirs(os.path.dirname(model_path_isolated), exist_ok=True)
    torch.save({
        'model_state_dict': agent_isolated.model.state_dict(),
        'optimizer_state_dict': agent_isolated.optimizer.state_dict(),
        'epsilon': agent_isolated.epsilon
    }, model_path_isolated)

    # 
    prioritized_metrics.end_training()

    # final samples
    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, _, sim, triggered in high_reward_samples:
            prioritized_metrics.record_final_output_sample(triggered, target_path)

    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, _, sim, triggered in high_reward_samples:
            prioritized_metrics.record_final_output_sample(triggered, target_path)

    # Excel
    performance_excel_path = os.path.join(results_save_dir, "Metric_.xlsx")
    samples_excel_path = os.path.join(results_save_dir, "final samples_.xlsx")

    append_performance_metrics_to_excel(prioritized_metrics, performance_excel_path, run_number)
    append_final_samples_to_excel(agent, agent_isolated, similar_group, isolated_group, targetPaths, samples_excel_path,
                                  run_number)

    #  runMetric
    avg_similarity = np.mean(
        prioritized_metrics.final_output_similarities) if prioritized_metrics.final_output_similarities else 0
    training_time = prioritized_metrics.end_time - prioritized_metrics.start_time
    avg_priority = np.mean(prioritized_metrics.priority_statistics) if prioritized_metrics.priority_statistics else 0
    print(f"\nRun  {run_number}  runcompleted:")
    print(f"  Average Similarity: {avg_similarity:.4f}")
    print(f"  Training Time: {training_time:.2f} seconds")
    print(f"  : {prioritized_metrics.step_count}")
    print(f"  : {avg_priority:.4f}")


if __name__ == "__main__":
    results_save_dir = r"D:\Experiment\CNN\DQNNEW\results\prioritized_results"
    os.makedirs(results_save_dir, exist_ok=True)

    # 20
    NUM_RUNS = 20

    print("=" * 80)
    print(f" {NUM_RUNS} DQN")
    print("=" * 80)

    for run in range(1, NUM_RUNS + 1):
        try:
            run_single_experiment(run, results_save_dir)
        except Exception as e:
            print(f"\nRun  {run}  run: {str(e)}")
            import traceback

            traceback.print_exc()
            continue

    print("\n" + "=" * 80)
    print(f" {NUM_RUNS}  runcompleted")
    print(f": {results_save_dir}")
    print("=" * 80)