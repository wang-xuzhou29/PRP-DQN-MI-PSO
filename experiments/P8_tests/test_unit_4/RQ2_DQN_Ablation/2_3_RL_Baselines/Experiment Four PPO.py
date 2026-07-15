
import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
import numpy as np
import random
import time
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
from datetime import datetime

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# ===  ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'ACTION_DIM': 3,
    'MIN_VALUES': np.array([1, 1, 1]),
    'MAX_VALUES': np.array([128, 200, 255]),
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 3,
    'PPO_BATCH_SIZE': 64,
    'PPO_EPOCHS': 4,
    'CLIP_EPSILON': 0.2,
    'GAMMA': 0.99,
    'GAE_LAMBDA': 0.95,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'ACTOR_LR': 3e-4,
    'CRITIC_LR': 3e-4,
    'NUM_RUNS': 20,
    'TOP_K_SAMPLES': 20,
    'TARGET_PATHS': [
        {1, 2, 4, 8, 15, 20, 35, 41, 46, 71, 73, 76, 77, 78, 82, 83, 84, 87, 97, 100, 109, 110, 116, 118, 121, 122, 123,
         125, 126, 131, 132, 133, 134, 137, 141, 144, 155, 156},
        {1, 2, 4, 8, 15, 20, 35, 46, 71, 73, 76, 77, 82, 83, 84, 87, 90, 91, 109, 110, 116, 118, 121, 122, 123, 125,
         126, 131, 132, 133, 134, 137, 140, 141, 144, 155, 156},
        {1, 2, 4, 8, 15, 20, 35, 46, 71, 73, 76, 77, 82, 83, 84, 87, 92, 97, 100, 109, 110, 116, 118, 121, 122, 123,
         125, 126, 131, 132, 133, 134, 137, 141, 144, 155, 156},
        {1, 2, 4, 8, 15, 20, 35, 46, 71, 73, 76, 77, 78, 82, 83, 84, 87, 90, 97, 100, 109, 110, 114, 115, 117, 119, 120,
         124, 127, 128, 134, 137, 140, 141, 144, 155, 156},
        {2, 8, 9, 24, 28, 35, 46, 71, 73, 76, 77, 78, 84, 87, 88, 90, 91, 100, 109, 110, 115, 117, 118, 128, 134, 135,
         141, 144, 146, 147, 148, 150, 151, 153, 154, 155},
        {24, 26, 27, 28, 29, 30, 37, 49, 52, 53, 55, 56, 57, 58, 61, 62, 63, 67, 70, 73, 74, 100, 110, 111, 112, 116,
         117, 126, 129, 130, 131, 141, 142, 144, 155, 156},
        {24, 26, 28, 29, 37, 49, 50, 52, 53, 55, 56, 57, 58, 61, 62, 63, 67, 70, 73, 74, 100, 110, 111, 116, 117, 126,
         129, 130, 131, 134, 141, 142, 144, 155, 156},
        {1, 2, 4, 8, 15, 20, 35, 41, 46, 71, 73, 76, 77, 78, 82, 83, 84, 87, 90, 97, 100, 109, 110, 114, 115, 117, 119,
         120, 124, 127, 128, 134, 137, 152, 157},
        {25, 31, 33, 34, 35, 36, 46, 49, 52, 53, 55, 57, 58, 61, 62, 63, 67, 70, 73, 74, 100, 110, 111, 116, 117, 126,
         129, 130, 131, 141, 142, 144, 155, 156},
        {1, 2, 7, 8, 9, 10, 24, 28, 35, 46, 54, 60, 71, 73, 81, 90, 91, 100, 109, 110, 117, 118, 128, 134, 141, 144,
         146, 147, 148, 150, 151, 153, 154, 155},
        {14, 16, 17, 19, 22, 25, 35, 45, 46, 60, 61, 65, 69, 76, 77, 82, 84, 87, 90, 93, 97, 110, 111, 112, 116, 118,
         126, 129, 130, 131, 144, 155, 156},
        {14, 16, 17, 19, 37, 39, 40, 41, 42, 43, 44, 47, 60, 61, 65, 69, 76, 77, 82, 84, 87, 90, 97, 116, 118, 126, 129,
         131, 134, 137, 144, 155, 156},
        {14, 16, 17, 19, 37, 39, 40, 41, 42, 43, 44, 47, 60, 61, 65, 69, 76, 77, 82, 84, 87, 94, 97, 116, 118, 126, 129,
         131, 134, 137, 144, 155, 156},
        {24, 26, 27, 28, 29, 30, 32, 37, 49, 52, 53, 55, 56, 57, 58, 61, 62, 63, 67, 70, 73, 74, 96, 97, 100, 110, 112,
         128, 141, 142, 144, 155, 156},
        {1, 3, 4, 5, 10, 13, 15, 20, 21, 22, 25, 35, 46, 76, 77, 97, 99, 100, 110, 111, 112, 116, 118, 126, 129, 130,
         131, 141, 142, 144, 155, 156},
        {14, 16, 17, 19, 25, 35, 38, 45, 46, 60, 61, 65, 69, 76, 77, 82, 84, 87, 90, 93, 97, 110, 111, 116, 118, 126,
         129, 130, 131, 144, 155, 156},
        {14, 16, 17, 18, 19, 22, 23, 25, 35, 46, 60, 65, 69, 76, 77, 82, 84, 87, 90, 93, 97, 100, 101, 104, 107, 110,
         111, 112, 128, 144, 155, 156},
        {1, 2, 6, 7, 8, 11, 12, 14, 24, 26, 27, 28, 29, 30, 32, 37, 97, 99, 100, 110, 111, 112, 116, 117, 126, 129, 130,
         131, 134, 141, 142, 144},
        {1, 4, 5, 13, 15, 20, 35, 46, 60, 61, 65, 69, 76, 77, 82, 84, 87, 94, 97, 100, 101, 104, 106, 109, 110, 128,
         134, 137, 144, 155, 156},
        {25, 29, 35, 38, 39, 46, 60, 61, 64, 65, 66, 68, 69, 73, 74, 87, 90, 93, 100, 110, 111, 116, 117, 118, 126, 129,
         131, 144, 155, 156},
        {1, 2, 8, 9, 10, 13, 20, 35, 46, 71, 73, 76, 77, 78, 84, 87, 88, 90, 91, 97, 98, 100, 109, 110, 136, 138, 141,
         142, 144, 155, 156},
        {1, 2, 8, 20, 35, 46, 71, 73, 81, 89, 90, 91, 100, 109, 110, 114, 115, 117, 119, 120, 124, 128, 134, 137, 140,
         141, 144, 155, 156},
        {1, 2, 8, 20, 35, 46, 71, 73, 76, 77, 78, 90, 91, 97, 100, 109, 110, 114, 115, 117, 119, 120, 124, 128, 134,
         137, 143, 145, 152},
        {2, 8, 9, 24, 35, 46, 71, 73, 76, 77, 78, 84, 87, 88, 90, 91, 100, 109, 110, 114, 115, 117, 118, 128, 138, 143,
         145, 149, 152},
        {1, 2, 4, 8, 13, 15, 20, 35, 46, 71, 73, 76, 77, 78, 82, 83, 84, 87, 92, 94, 102, 105, 107, 108, 128, 134, 137,
         144, 155, 156},
        {1, 2, 4, 8, 13, 15, 20, 35, 46, 71, 73, 76, 77, 82, 83, 84, 87, 92, 94, 95, 97, 100, 109, 110, 128, 134, 137,
         144, 155, 156},
        {1, 4, 5, 13, 15, 20, 35, 46, 60, 61, 65, 69, 76, 77, 82, 84, 87, 94, 102, 103, 105, 107, 108, 128, 134, 137,
         144, 155, 156},
        {3, 5, 6, 9, 11, 12, 14, 35, 46, 48, 51, 52, 53, 56, 58, 64, 66, 67, 70, 72, 76, 79, 84, 85, 87, 90, 93, 108,
         144, 155, 156},
        {1, 3, 4, 5, 6, 11, 12, 14, 24, 30, 31, 33, 35, 36, 46, 54, 59, 60, 71, 73, 76, 79, 84, 108, 134, 141, 142, 144,
         155, 156},
        {3, 5, 6, 9, 11, 12, 14, 35, 46, 60, 61, 65, 69, 71, 73, 76, 77, 78, 84, 85, 87, 90, 93, 108, 113, 128, 144,
         155, 156},
        {1, 3, 4, 5, 6, 11, 12, 14, 24, 30, 35, 36, 46, 54, 59, 60, 71, 73, 80, 81, 108, 113, 134, 141, 142, 144, 155,
         156},
        {1, 2, 4, 7, 8, 10, 13, 15, 20, 35, 46, 71, 73, 81, 86, 92, 94, 97, 98, 100, 109, 110, 134, 137, 144, 155, 156},
        {1, 6, 11, 14, 24, 25, 26, 28, 29, 37, 48, 51, 54, 58, 59, 60, 75, 84, 108, 128, 134, 141, 142, 144, 155, 156},
        {2, 7, 8, 9, 10, 24, 28, 35, 46, 71, 73, 81, 90, 91, 100, 109, 110, 117, 118, 128, 139, 141, 144}
    ],
}

# ===  ===
def execute_Tr(a):
    """
    替换原有的编排规则函数
    参数a: 包含3个元素的元组或数组，分别对应path_depth, file_count, access_level
    返回: 触发的规则编号集合
    """
    path_depth, file_count, access_level = float(a[0]), int(a[1]), float(a[2])

    # 使用正确的变量名
    x, y, z = path_depth, file_count, access_level

    triggered = set()
    # 创建一个字典来存储b数组的值，用于跟踪哪些规则被触发
    b = {}
    x_score = max(0, 100 - x * 1.67)

    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((y * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[0] = 1
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((z * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[1] = 2
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * x) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[2] = 3
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[3] = 4
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * 43) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[4] = 5
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((74 * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[5] = 6
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (47 + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[6] = 7
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (x + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[7] = 8
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (y + 1) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[8] = 9
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 10) > 50 and x ** 2 + y ** 2 > z ** 2):
        b[9] = 10
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) != 50 and x ** 2 + y ** 2 > z ** 2):
        b[10] = 11
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 29 and x ** 2 + y ** 2 > z ** 2):
        b[11] = 12
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 69 and x ** 2 + y ** 2 > z ** 2):
        b[12] = 13
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 or x ** 2 + y ** 2 > z ** 2): b[
        13] = 14
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and y ** 2 + y ** 2 > z ** 2):
        b[14] = 15
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 3 + y ** 2 > z ** 2):
        b[15] = 16
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + x ** 2 > z ** 2):
        b[16] = 17
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 3 > z ** 2):
        b[17] = 18
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 != z ** 2):
        b[18] = 19
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2.5):
        b[19] = 20
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + y ** 1.4 > z ** 2):
        b[20] = 21
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and x ** 2 + 34 ** 2 > z ** 2):
        b[21] = 22
    if ((x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2) != ((x * z) / (z + 1) > 50 and 64 ** 2 + y ** 2 > z ** 2):
        b[22] = 23
    if (x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > z ** 2:
        pattern_type = "311A1"
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (z ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[23] = 24
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - z ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[24] = 25
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 1.5) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[25] = 26
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2.3 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[26] = 27
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (45 ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[27] = 28
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - 31 ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[28] = 29
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[29] = 30
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (x + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[30] = 31
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 2) < -30 or (abs(x - y) * z) / 100 > 45): b[31] = 32
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (26 + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[32] = 33
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y - 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[33] = 34
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) != -30 or (abs(x - y) * z) / 100 > 45): b[34] = 35
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -10 or (abs(x - y) * z) / 100 > 45): b[35] = 36
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 and (abs(x - y) * z) / 100 > 45): b[36] = 37
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(z - y) * z) / 100 > 45): b[37] = 38
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - z) * z) / 100 > 45): b[38] = 39
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * y) / 100 > 45): b[39] = 40
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * x) / 100 > 45): b[40] = 41
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * 34) / 100 > 45): b[41] = 42
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - 26) * z) / 100 > 45): b[42] = 43
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(52 - y) * z) / 100 > 45): b[43] = 44
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 76 > 45): b[44] = 45
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 != 45): b[45] = 46
    if ((x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 55): b[46] = 47
    if (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * z) / 100 > 45:
        pattern_type = "311A2"

    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + x) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35): b[47] = 48
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((y + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35): b[48] = 49
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((24 + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35): b[49] = 50
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + 31) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35): b[50] = 51
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 1.5) ** 2 < z * 20 and (x * x * z) / 1000 < 35): b[51] = 52
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2.4 < z * 20 and (x * x * z) / 1000 < 35): b[52] = 53
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 != z * 20 and (x * x * z) / 1000 < 35): b[53] = 54
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < 25 * 20 and (x * x * z) / 1000 < 35): b[54] = 55
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 14 and (x * x * z) / 1000 < 35): b[55] = 56
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < x * 20 and (x * x * z) / 1000 < 35): b[56] = 57
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < y * 20 and (x * x * z) / 1000 < 35): b[57] = 58
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 1.2 < z * 20 and (x * x * z) / 1000 < 35): b[58] = 59
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 or (x * x * z) / 1000 < 35): b[59] = 60
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (y * x * z) / 1000 < 35): b[60] = 61
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (z * x * z) / 1000 < 35): b[61] = 62
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * 35 * z) / 1000 < 35): b[62] = 63
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * x * x) / 1000 < 35): b[63] = 64
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * x * y) / 1000 < 35): b[64] = 65
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * x * 26) / 1000 < 35): b[65] = 66
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * 54 * z) / 1000 < 35): b[66] = 67
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1050 < 35): b[67] = 68
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 != 35): b[68] = 69
    if (((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35) != (
            ((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 25): b[69] = 70
    if ((x + y) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35:
        pattern_type = "311A3"
    if (x < 40 and z < 50) != (x != 40 and z < 50): b[70] = 71
    if (x < 40 and z < 50) != (x < 29 and z < 50): b[71] = 72
    if (x < 40 and z < 50) != (x < 40 or z < 50): b[72] = 73
    if (x < 40 and z < 50) != (x < 40 and z != 50): b[73] = 74
    if (x < 40 and z < 50) != (x < 40 and z < 36): b[74] = 75
    if x < 40 and z < 50:
        pattern_type = 322
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 or x > 50) or (z < 40 and x > 55 and y < 40)): b[75] = 76
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z != 35 and x > 50) or (z < 40 and x > 55 and y < 40)): b[76] = 77
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 45 and x > 50) or (z < 40 and x > 55 and y < 40)): b[77] = 78
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x != 50) or (z < 40 and x > 55 and y < 40)): b[78] = 79
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 59) or (z < 40 and x > 55 and y < 40)): b[79] = 80
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) and (z < 40 and x > 55 and y < 40)): b[80] = 81
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) or (z != 40 and x > 55 and y < 40)): b[81] = 82
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) or (z < 48 and x > 55 and y < 40)): b[82] = 83
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) or (z < 40 or x > 55 and y < 40)): b[83] = 84
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) or (z < 40 and x != 55 and y < 40)): b[84] = 85
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) or (z < 40 and x > 65 and y < 40)): b[85] = 86
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) or (z < 40 and x > 55 or y < 40)): b[86] = 87
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) or (z < 40 and x > 55 and y != 40)): b[87] = 88
    if ((z < 35 and x > 50) or (z < 40 and x > 55 and y < 40)) != (
            (z < 35 and x > 50) or (z < 40 and x > 55 and y < 29)): b[88] = 89
    if (z < 35 and x > 50) or (z < 40 and x > 55 and y < 40):
        pattern_type = 323
    if (y < 25 and x > 60) != (y < 25 or x > 60): b[89] = 90
    if (y < 25 and x > 60) != (y != 25 and x > 60): b[90] = 91
    if (y < 25 and x > 60) != (y < 17 and x > 60): b[91] = 92
    if (y < 25 and x > 60) != (y < 25 and x != 60): b[92] = 93
    if (y < 25 and x > 60) != (y < 25 and x > 69): b[93] = 94
    if y < 25 and x > 60:
        pattern_type = 324
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 75 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)): b[94] = 95
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (22 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)): b[95] = 96
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 or 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)): b[96] = 97
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 29 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)): b[97] = 98
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 78 and (x * z) / 100 < 35) or (x < 60 and z < 60)): b[98] = 99
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 or (x * z) / 100 < 35) or (x < 60 and z < 60)): b[99] = 100
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (56 * z) / 100 < 35) or (x < 60 and z < 60)): b[100] = 101
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * 74) / 100 < 35) or (x < 60 and z < 60)): b[101] = 102
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 89 < 35) or (x < 60 and z < 60)): b[102] = 103
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 != 35) or (x < 60 and z < 60)): b[103] = 104
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 21) or (x < 60 and z < 60)): b[104] = 105
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (z * z) / 100 < 35) or (x < 60 and z < 60)): b[105] = 106
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * x) / 100 < 35) or (x < 60 and z < 60)): b[106] = 107
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) and (x < 60 and z < 60)): b[107] = 108
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x != 60 and z < 60)): b[108] = 109
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 or z < 60)): b[109] = 110
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z != 60)): b[110] = 111
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 68)): b[111] = 112
    if ((40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60)) != (
            (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 51 and z < 60)): b[112] = 113
    if (40 <= x < 65 and 40 <= z < 65 and (x * z) / 100 < 35) or (x < 60 and z < 60):
        pattern_type = 325
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((53 ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)): b[113] = 114
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 2.3) / (x + y + 1) < 20) or (z < 65 and x < 75)): b[114] = 115
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 1.4) / (x + y + 1) < 20) or (z < 65 and x < 75)): b[115] = 116
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((x ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)):
        b[116] = 117
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((y ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)):
        b[117] = 118
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((z ** 2) / (y + y + 1) < 20) or (z < 65 and x < 75)):
        b[118] = 119
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((z ** 2) / (z + y + 1) < 20) or (z < 65 and x < 75)):
        b[119] = 120
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((z ** 2) / (x + x + 1) < 20) or (z < 65 and x < 75)):
        b[120] = 121
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((z ** 2) / (x + z + 1) < 20) or (z < 65 and x < 75)):
        b[121] = 122
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 2) / (x + 45 + 1) < 20) or (z < 65 and x < 75)): b[122] = 123
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 2) / (33 + y + 1) < 20) or (z < 65 and x < 75)): b[123] = 124
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 2) / (x + y + 11) < 20) or (z < 65 and x < 75)): b[124] = 125
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 2) / (x + y + 1) != 20) or (z < 65 and x < 75)): b[125] = 126
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((z ** 2) / (x + y + 1) < 15) or (z < 65 and x < 75)):
        b[126] = 127
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 2) / (x + y + 1) < 20) and (z < 65 and x < 75)): b[127] = 128
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 2) / (x + y + 1) < 20) or (z != 65 and x < 75)): b[128] = 129
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((z ** 2) / (x + y + 1) < 20) or (z < 75 and x < 75)):
        b[129] = 130
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((z ** 2) / (x + y + 1) < 20) or (z < 65 or x < 75)):
        b[130] = 131
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (
            ((z ** 2) / (x + y + 1) < 20) or (z < 65 and x != 75)): b[131] = 132
    if (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75)) != (((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 79)):
        b[132] = 133
    if ((z ** 2) / (x + y + 1) < 20) or (z < 65 and x < 75):
        pattern_type = 326
    if (60 <= x < 82 and y > 50) != (60 <= x < 82 or y > 50): b[133] = 134
    if (60 <= x < 82 and y > 50) != (60 <= x < 89 and y > 50): b[134] = 135
    if (60 <= x < 82 and y > 50) != (67 <= x < 82 and y > 50): b[135] = 136
    if (60 <= x < 82 and y > 50) != (60 <= x < 82 and y != 50): b[136] = 137
    if (60 <= x < 82 and y > 50) != (60 <= x < 82 and y > 62): b[137] = 138
    if 60 <= x < 82 and y > 50:
        pattern_type = 327
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 77 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)): b[138] = 139
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (22 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)): b[139] = 140
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 or x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)): b[140] = 141
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x != 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)): b[141] = 142
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 85 and (y * x) / 100 < 52) or (y < 60 and x > 80)): b[142] = 143
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 or (y * x) / 100 < 52) or (y < 60 and x > 80)): b[143] = 144
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (x * x) / 100 < 52) or (y < 60 and x > 80)): b[144] = 145
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * y) / 100 < 52) or (y < 60 and x > 80)): b[145] = 146
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (43 * x) / 100 < 52) or (y < 60 and x > 80)): b[146] = 147
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * 31) / 100 < 52) or (y < 60 and x > 80)): b[147] = 148
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 78 < 52) or (y < 60 and x > 80)): b[148] = 149
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 100 != 52) or (y < 60 and x > 80)): b[149] = 150
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 100 < 62) or (y < 60 and x > 80)): b[150] = 151
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 100 < 52) and (y < 60 and x > 80)): b[151] = 152
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y != 60 and x > 80)): b[152] = 153
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 65 and x > 80)): b[153] = 154
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 or x > 80)): b[154] = 155
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x != 80)): b[155] = 156
    if ((40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80)) != (
            (40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 88)): b[156] = 157
    if (40 <= y < 65 and x > 75 and (y * x) / 100 < 52) or (y < 60 and x > 80):
        pattern_type = 328

    # 返回被触发的规则编号集合
    return set(b.values())

# ===  ===
# === 状态处理辅助函数 ===
def clip_state(state):
    return np.clip(state, EXPERIMENT_CONFIG['MIN_VALUES'], EXPERIMENT_CONFIG['MAX_VALUES'])

def denormalize_state(normalized_state):
    """将归一化状态还原为原始状态"""
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return normalized_state * (max_vals - min_vals) / 2 + (min_vals + max_vals) / 2

def coverage_similarity(triggered, target_path):
    """
    Similarity:  / target paths
    
    """
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0

    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)

def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']

    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']

    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']

    return reward

# === PPO Actor ===
class PPOActor(nn.Module):
    def __init__(self, state_dim=3, action_dim=3, hidden_dim=256):
        super(PPOActor, self).__init__()

        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)

        self.mean_head = nn.Linear(hidden_dim, action_dim)
        self.log_std_head = nn.Linear(hidden_dim, action_dim)

        self.action_scale = 10.0

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))
        x = F.relu(self.fc3(x))

        mean = torch.tanh(self.mean_head(x)) * self.action_scale
        log_std = torch.clamp(self.log_std_head(x), -20, 2)

        return mean, log_std

    def get_action_and_log_prob(self, state):
        mean, log_std = self.forward(state)
        std = torch.exp(log_std)

        dist = torch.distributions.Normal(mean, std)
        action = dist.sample()
        log_prob = dist.log_prob(action).sum(dim=-1)

        return action, log_prob

    def get_log_prob(self, state, action):
        mean, log_std = self.forward(state)
        std = torch.exp(log_std)

        dist = torch.distributions.Normal(mean, std)
        log_prob = dist.log_prob(action).sum(dim=-1)

        return log_prob

# === PPO Critic ===
class PPOCritic(nn.Module):
    def __init__(self, state_dim=3, hidden_dim=256):
        super(PPOCritic, self).__init__()

        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)
        self.value_head = nn.Linear(hidden_dim, 1)

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))
        x = F.relu(self.fc3(x))
        value = self.value_head(x)
        return value.squeeze(-1)

# === PPO ===
class PPOBuffer:
    def __init__(self):
        self.states = []
        self.actions = []
        self.rewards = []
        self.values = []
        self.log_probs = []
        self.dones = []
        self.path_indices = []
        self.similarities = []

    def store(self, state, action, reward, value, log_prob, done, path_idx, similarity):
        self.states.append(state)
        self.actions.append(action)
        self.rewards.append(reward)
        self.values.append(value)
        self.log_probs.append(log_prob)
        self.dones.append(done)
        self.path_indices.append(path_idx)
        self.similarities.append(similarity)

    def compute_advantages(self):
        rewards = np.array(self.rewards)
        values = np.array(self.values)
        dones = np.array(self.dones)

        advantages = np.zeros_like(rewards)
        returns = np.zeros_like(rewards)
        last_advantage = 0

        for t in reversed(range(len(rewards))):
            if t == len(rewards) - 1:
                next_value = 0
            else:
                next_value = values[t + 1]

            delta = rewards[t] + EXPERIMENT_CONFIG['GAMMA'] * next_value * (1 - dones[t]) - values[t]
            advantages[t] = delta + EXPERIMENT_CONFIG['GAMMA'] * EXPERIMENT_CONFIG['GAE_LAMBDA'] * (
                    1 - dones[t]) * last_advantage
            last_advantage = advantages[t]

        returns = advantages + values
        advantages = (advantages - np.mean(advantages)) / (np.std(advantages) + 1e-8)

        return advantages, returns

    def get_batch_iterator(self, advantages, returns):
        indices = np.arange(len(self.states))
        np.random.shuffle(indices)

        batch_size = EXPERIMENT_CONFIG['PPO_BATCH_SIZE']
        for start in range(0, len(self.states), batch_size):
            end = min(start + batch_size, len(self.states))
            batch_indices = indices[start:end]

            yield {
                'states': torch.FloatTensor(np.array([self.states[i] for i in batch_indices])).to(device),
                'actions': torch.FloatTensor(np.array([self.actions[i] for i in batch_indices])).to(device),
                'old_log_probs': torch.FloatTensor(np.array([self.log_probs[i] for i in batch_indices])).to(device),
                'advantages': torch.FloatTensor(advantages[batch_indices]).to(device),
                'returns': torch.FloatTensor(returns[batch_indices]).to(device)
            }

    def get_top_k_per_path(self, num_paths, k=20):
        path_samples = {i: [] for i in range(num_paths)}

        for idx in range(len(self.states)):
            path_idx = self.path_indices[idx]
            similarity = self.similarities[idx]
            state = self.states[idx]
            path_samples[path_idx].append((idx, similarity, state))

        top_k_results = {}
        for path_idx in range(num_paths):
            samples = path_samples[path_idx]
            if len(samples) == 0:
                top_k_results[path_idx] = []
                continue

            samples.sort(key=lambda x: x[1], reverse=True)
            top_k = samples[:k]

            top_k_results[path_idx] = []
            for sample in top_k:
                normalized_state = sample[2]
                original_state = denormalize_state(normalized_state)
                original_state_int = np.round(original_state).astype(int)

                top_k_results[path_idx].append({
                    'state': original_state_int,
                    'similarity': sample[1],
                    'triggered': execute_Tr(*original_state_int)
                })

        return top_k_results

    def clear(self):
        self.states.clear()
        self.actions.clear()
        self.rewards.clear()
        self.values.clear()
        self.log_probs.clear()
        self.dones.clear()
        self.path_indices.clear()
        self.similarities.clear()

    def __len__(self):
        return len(self.states)

# === PPO ===
class PPOAgent:
    def __init__(self, state_dim=3, action_dim=3):
        self.state_dim = state_dim
        self.action_dim = action_dim

        self.actor = PPOActor(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.actor_optimizer = optim.Adam(self.actor.parameters(), lr=EXPERIMENT_CONFIG['ACTOR_LR'])

        self.critic = PPOCritic(state_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_optimizer = optim.Adam(self.critic.parameters(), lr=EXPERIMENT_CONFIG['CRITIC_LR'])

        self.buffer = PPOBuffer()
        self.update_count = 0

    def get_action(self, state):
            min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
            max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
            normalized_state = (state - (min_vals + max_vals) / 2) / ((max_vals - min_vals) / 2)
            state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)

            with torch.no_grad():
                action, log_prob = self.actor.get_action_and_log_prob(state_tensor)
                value = self.critic(state_tensor)

            action = action.cpu().numpy()[0]
            log_prob = log_prob.cpu().item()
            value = value.cpu().item()

            return action, log_prob, value
    def store_experience(self, state, action, reward, value, log_prob, done, path_idx, similarity):
            min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
            max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
            normalized_state = (state - (min_vals + max_vals) / 2) / ((max_vals - min_vals) / 2)
            self.buffer.store(normalized_state, action, reward, value, log_prob, done, path_idx, similarity)

    def update(self):
        if len(self.buffer) == 0:
            return

        advantages, returns = self.buffer.compute_advantages()

        for epoch in range(EXPERIMENT_CONFIG['PPO_EPOCHS']):
            for batch in self.buffer.get_batch_iterator(advantages, returns):
                new_log_probs = self.actor.get_log_prob(batch['states'], batch['actions'])
                ratio = torch.exp(new_log_probs - batch['old_log_probs'])

                surr1 = ratio * batch['advantages']
                surr2 = torch.clamp(ratio, 1 - EXPERIMENT_CONFIG['CLIP_EPSILON'],
                                    1 + EXPERIMENT_CONFIG['CLIP_EPSILON']) * batch['advantages']
                actor_loss = -torch.min(surr1, surr2).mean()

                self.actor_optimizer.zero_grad()
                actor_loss.backward()
                torch.nn.utils.clip_grad_norm_(self.actor.parameters(), 0.5)
                self.actor_optimizer.step()

                new_values = self.critic(batch['states'])
                critic_loss = F.mse_loss(new_values, batch['returns'])

                self.critic_optimizer.zero_grad()
                critic_loss.backward()
                torch.nn.utils.clip_grad_norm_(self.critic.parameters(), 0.5)
                self.critic_optimizer.step()

        self.update_count += 1
        self.buffer.clear()

        if self.update_count % 2 == 0:
            print(f"  -> PPOcompleted (Run {self.update_count})")

# === Metric ===
def calculate_run_performance(run_idx, ppo_results, training_time, total_steps, update_count, agent):
    """ runMetric"""
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    # 1. (Total Reward)
    total_reward = 0
    # 2. (Average Reward)
    average_reward = 0
    # 5. (Convergence)
    convergence = 0
    # 12. (Environment Adaptability)
    environment_adaptability = 0
    # 13. (Generalization Ability)
    generalization_ability = 0
    # 15. (Computational Efficiency)
    computational_efficiency = 0
    # 16. (Policy Update Frequency)
    policy_update_frequency = 0

    # Similarity
    all_similarities = []

    # Metric
    total_samples = 0
    all_rewards = []

    for path_idx in range(num_paths):
        samples = ppo_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']

            total_reward += reward
            all_rewards.append(reward)
            all_similarities.append(similarity)
            total_samples += 1

    # 1. 
    total_reward = total_reward

    # 2. 
    if total_samples > 0:
        average_reward = total_reward / total_samples

    # 5. (Average Similarity)
    if all_similarities:
        convergence = np.mean(all_similarities)

    # 12. (Similarity)
    if len(all_similarities) > 1:
        environment_adaptability = 1 / (np.std(all_similarities) + 1e-8)

    # 13. (Average Similarity)
    generalization_ability = convergence

    # 15. (/ seconds)
    if training_time > 0:
        computational_efficiency = total_steps / training_time

    # 16. 
    if training_time > 0:
        policy_update_frequency = update_count / training_time

    # Similarity
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0

    return {
        'Run': run_idx + 1,

        # Metric
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),

        # Similarity
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }

# === Excel ===
def export_to_excel(all_ppo_results, all_performance_data, target_paths, output_path="PPO_20 run.xlsx"):
    """20 runPPOExcel"""
    print("\nExcel...")

    # 
    all_ppo_summary_data = []
    all_ppo_detailed_data = []

    #  run
    for run_idx, (ppo_results, performance_data) in enumerate(zip(all_ppo_results, all_performance_data)):
        # ===== Sheet1: PPOPath  =====
        ppo_summary_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = ppo_results[path_idx]

            if len(samples) == 0:
                ppo_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue

            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = '' if perfect_count > 0 else ''

            ppo_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })

        all_ppo_summary_data.extend(ppo_summary_data)

        # ===== Sheet2: PPODetailed Sample Data =====
        ppo_detailed_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = ppo_results[path_idx]

            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']

                ppo_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })

        all_ppo_detailed_data.extend(ppo_detailed_data)

    # Excel
    ppo_summary_df = pd.DataFrame(all_ppo_summary_data)
    ppo_detailed_df = pd.DataFrame(all_ppo_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet1: PPOPath 
        ppo_summary_df.to_excel(writer, sheet_name='PPOPath ', index=False)

        # Sheet2: PPODetailed Sample Data
        ppo_detailed_df.to_excel(writer, sheet_name='PPODetailed Sample Data', index=False)

        # Sheet3: Metric - 
        selected_columns = [
            'Run',
            'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df_selected = performance_df[selected_columns]
        performance_df_selected.to_excel(writer, sheet_name='Metric', index=False)

        # 
        workbook = writer.book

        # 
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        perfect_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # 

        # === Sheet1 ===
        ws1 = writer.sheets['PPOPath ']
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 
        for row_idx in range(2, ws1.max_row + 1):
            if ws1.cell(row_idx, 9).value == '':  # Run 9""
                for col_idx in range(1, ws1.max_column + 1):
                    ws1.cell(row_idx, col_idx).fill = perfect_fill

        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 15
        ws1.column_dimensions['F'].width = 15
        ws1.column_dimensions['G'].width = 15
        ws1.column_dimensions['H'].width = 15
        ws1.column_dimensions['I'].width = 15
        ws1.column_dimensions['J'].width = 50

        # === Sheet2 ===
        ws2 = writer.sheets['PPODetailed Sample Data']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 15
        ws2.column_dimensions['I'].width = 40
        ws2.column_dimensions['J'].width = 40
        ws2.column_dimensions['K'].width = 15
        ws2.column_dimensions['L'].width = 15

        # === Sheet3 ===
        ws3 = writer.sheets['Metric']
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        for col in columns:
            ws3.column_dimensions[col].width = 18

    print(f"Excel: {output_path}")
    print(f"  - Sheet1: PPOPath  ({len(all_ppo_summary_data)})")
    print(f"  - Sheet2: PPODetailed Sample Data ({len(all_ppo_detailed_data)})")
    print(f"  - Sheet3: Metric ({len(all_performance_data)})")

# === PPO ===
def train_ppo_workflow():
    print("=" * 80)
    print("PPO")
    print("Similarity:  / target paths")
    print("=" * 80)

    agent = PPOAgent()
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n: Path {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']}")
    path_samples = {}
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']

    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            # 分别生成 X, Y, Z 的随机整数
            state = np.array([
                np.random.randint(min_vals[0], max_vals[0] + 1),
                np.random.randint(min_vals[1], max_vals[1] + 1),
                np.random.randint(min_vals[2], max_vals[2] + 1)
            ], dtype=np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  Path  {path_idx + 1}/{num_paths}:  {len(samples)} ")

    # 
    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size

    print(f"\n: {batch_size},{EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(f": {num_batches} /Path  x {num_paths} Path  = {num_batches * num_paths} ")
    print("-" * 80)

    global_buffer = PPOBuffer()

    for batch_idx in range(num_batches):
        print(f"\n {batch_idx + 1}/{num_batches}")

        for path_idx in range(num_paths):
            target_path = target_paths[path_idx]
            batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]

            batch_rewards = []
            batch_similarities = []

            for sample_idx, initial_state in enumerate(batch_samples):
                state = initial_state.copy()
                episode_reward = 0
                final_similarity = 0

                for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                    action, log_prob, value = agent.get_action(state)

                    next_state = state + action
                    next_state = clip_state(next_state)

                    triggered = execute_Tr(*next_state)
                    reward = unified_reward_function(triggered, target_path)
                    similarity = coverage_similarity(triggered, target_path)

                    done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)

                    agent.store_experience(state, action, reward, value, log_prob, done, path_idx, similarity)
                    normalized_state = (state - (min_vals + max_vals) / 2) / ((max_vals - min_vals) / 2)
                    global_buffer.store(
                        normalized_state,
                        action, reward, value, log_prob, done, path_idx, similarity
                    )

                    state = next_state
                    episode_reward += reward
                    final_similarity = similarity
                    total_steps += 1

                batch_rewards.append(episode_reward)
                batch_similarities.append(final_similarity)

            avg_reward = np.mean(batch_rewards)
            avg_similarity = np.mean(batch_similarities)
            print(f"  Path {path_idx + 1}: ={avg_reward:.2f}, Average Similarity={avg_similarity:.4f}")

        print(f"\n  PPO...")
        agent.update()
        print(f"  : {len(global_buffer)}")

    training_time = time.time() - start_time

    print("\n" + "=" * 80)
    print(f"PPOcompleted! Total elapsed time: {training_time:.2f} seconds, : {total_steps}")
    print(f": {len(global_buffer)}")
    print(f"PPO: {agent.update_count}")
    print("=" * 80)

    # Top-K
    print(f"\nPath SimilarityMaximum{EXPERIMENT_CONFIG['TOP_K_SAMPLES']}...")
    ppo_top_k_results = global_buffer.get_top_k_per_path(num_paths, EXPERIMENT_CONFIG['TOP_K_SAMPLES'])

    return agent, ppo_top_k_results, training_time, total_steps, agent.update_count

# ===  ===
def main():
    print("\n" + "=" * 80)
    print("PPO - 20 run")
    print("Metric")
    print("=" * 80)

    all_ppo_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    # 20
    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'=' * 80}")
        print(f"Start run  {run_idx + 1}/{EXPERIMENT_CONFIG['NUM_RUNS']}  run")
        print(f"{'=' * 80}")

        # PPO
        ppo_agent, ppo_results, training_time, total_steps, update_count = train_ppo_workflow()

        # Metric
        performance_data = calculate_run_performance(
            run_idx, ppo_results, training_time, total_steps, update_count, ppo_agent
        )

        # 
        all_ppo_results.append(ppo_results)
        all_performance_data.append(performance_data)

        print(f"\nRun  {run_idx + 1}  runcompleted!")
        print(f"  Total Reward: {performance_data['Total Reward']}")
        print(f"  Average Reward: {performance_data['Average Reward']}")
        print(f"  Average Similarity: {performance_data['Average Similarity']}")

    # Excel(20 run)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"PPO_20 run_{timestamp}.xlsx"
    export_to_excel(all_ppo_results, all_performance_data, target_paths, output_path)

    # 
    print("\n" + "=" * 80)
    print("20 run")
    print("=" * 80)

    # Metric
    total_rewards = [p['Total Reward'] for p in all_performance_data]
    average_rewards = [p['Average Reward'] for p in all_performance_data]
    convergences = [p['Convergence'] for p in all_performance_data]
    environment_adaptabilities = [p['Environment Adaptability'] for p in all_performance_data]
    generalization_abilities = [p['Generalization Ability'] for p in all_performance_data]
    computational_efficiencies = [p['Computational Efficiency'] for p in all_performance_data]
    policy_update_frequencies = [p['Policy Update Frequency'] for p in all_performance_data]
    avg_similarities = [p['Average Similarity'] for p in all_performance_data]

    print(f"\nTotal Reward Statistics:")
    print(f"  Mean: {np.mean(total_rewards):.2f}")
    print(f"  Standard deviation: {np.std(total_rewards):.2f}")

    print(f"\nAverage Reward Statistics:")
    print(f"  Mean: {np.mean(average_rewards):.4f}")
    print(f"  Standard deviation: {np.std(average_rewards):.4f}")

    print(f"\nConvergence Statistics:")
    print(f"  Mean: {np.mean(convergences):.4f}")
    print(f"  Standard deviation: {np.std(convergences):.4f}")

    print(f"\nEnvironment Adaptability Statistics:")
    print(f"  Mean: {np.mean(environment_adaptabilities):.4f}")
    print(f"  Standard deviation: {np.std(environment_adaptabilities):.4f}")

    print(f"\nGeneralization Ability Statistics:")
    print(f"  Mean: {np.mean(generalization_abilities):.4f}")
    print(f"  Standard deviation: {np.std(generalization_abilities):.4f}")

    print(f"\nComputational Efficiency Statistics:")
    print(f"  Mean: {np.mean(computational_efficiencies):.2f}")
    print(f"  Standard deviation: {np.std(computational_efficiencies):.2f}")

    print(f"\nPolicy Update Frequency Statistics:")
    print(f"  Mean: {np.mean(policy_update_frequencies):.4f}")
    print(f"  Standard deviation: {np.std(policy_update_frequencies):.4f}")

    print("\n" + "=" * 80)
    print(f" {EXPERIMENT_CONFIG['NUM_RUNS']} completed!")
    print("=" * 80)

if __name__ == "__main__":
    main()