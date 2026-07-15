import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

def execute_validation_rules(a):
    """
    替换原有的编排规则函数
    参数a: 包含3个元素的元组或数组，分别对应path_depth, file_count, access_level
    返回: 触发的规则编号集合
    """
    path_depth, file_count, access_level = float(a[0]), int(a[1]), float(a[2])

    # 使用正确的变量名
    x, y, z = path_depth, file_count, access_level

    triggered = set()
    # 创建一个字典来存储b数组的值，用于跟踪哪些规则被触发
    b = {}
    x_score = max(0, 100 - x * 1.67)

    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (y * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4): b[0] = 1
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (z * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4): b[1] = 2
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (43 * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4): b[2] = 3
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (x * 51) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4): b[3] = 4
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (x * z) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4): b[4] = 5
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (x * x) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4): b[5] = 6
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (x * y) / (26 + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4): b[6] = 7
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (x * y) / (z + 8) > 50 and x ** 2 + y ** 2 > x ** 2.4): b[7] = 8
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (x * y) / (x + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4): b[8] = 9
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (x * y) / (y + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4): b[9] = 10
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (x * y) / (z + 1) != 50 and x ** 2 + y ** 2 > x ** 2.4): b[10] = 11
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (x * y) / (z + 1) > 41 and x ** 2 + y ** 2 > x ** 2.4): b[11] = 12
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (x * y) / (z + 1) > 62 and x ** 2 + y ** 2 > x ** 2.4): b[12] = 13
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (x * y) / (z + 1) > 50 or x ** 2 + y ** 2 > x ** 2.4): b[13] = 14
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (x * y) / (z + 1) > 50 and y ** 2 + y ** 2 > x ** 2.4): b[14] = 15
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (x * y) / (z + 1) > 50 and z ** 2 + y ** 2 > x ** 2.4): b[15] = 16
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (x * y) / (z + 1) > 50 and x ** 2 + x ** 2 > x ** 2.4): b[16] = 17
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (x * y) / (z + 1) > 50 and x ** 2 + z ** 2 > x ** 2.4): b[17] = 18
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (x * y) / (z + 1) > 50 and x ** 2 + y ** 2.1 > x ** 2.4): b[18] = 19
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 3):
        b[19] = 20
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (x * y) / (z + 1) > 50 and 24 ** 2 + y ** 2 > x ** 2.4): b[20] = 21
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (x * y) / (z + 1) > 50 and x ** 2 + 33 ** 2 > x ** 2.4): b[21] = 22
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > 53 ** 2.4): b[22] = 23
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (x * y) / (z + 1) > 50 and x ** 2 + y ** 2 != x ** 2.4): b[23] = 24
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (x * y) / (z + 1) > 50 and x ** 2.2 + y ** 2 > x ** 2.4): b[24] = 25
    if ((x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4) != (
            (x * y) / (z + 1) > 50 and x ** 2 + y ** 1.7 > x ** 2.4): b[25] = 26
    if (x * y) / (z + 1) > 50 and x ** 2 + y ** 2 > x ** 2.4:
        pattern_type = "111A1"
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (x ** 1.7 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45): b[26] = 27
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (x ** 2 - y ** 1.5) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45): b[27] = 28
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (z ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45): b[28] = 29
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (43 ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45): b[29] = 30
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (x ** 2 - z ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45): b[30] = 31
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (x ** 2 - 27 ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45): b[31] = 32
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (x ** 2 - y ** 2) / (x + 0.1) < -30 or (abs(x - y) * x) / 100 > 45): b[32] = 33
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (x ** 2 - y ** 2) / (y + 0.1) < -30 or (abs(x - y) * x) / 100 > 45): b[33] = 34
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 1.3) < -30 or (abs(x - y) * x) / 100 > 45): b[34] = 35
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) != -30 or (abs(x - y) * x) / 100 > 45): b[35] = 36
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -15 or (abs(x - y) * x) / 100 > 45): b[36] = 37
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (x ** 2 - y ** 2) / (44 + 0.1) < -30 or (abs(x - y) * x) / 100 > 45): b[37] = 38
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 and (abs(x - y) * x) / 100 > 45): b[38] = 39
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(52 - y) * x) / 100 > 45): b[39] = 40
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - 33) * x) / 100 > 45): b[40] = 41
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * 24) / 100 > 45): b[41] = 42
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(z - y) * x) / 100 > 45): b[42] = 43
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - z) * x) / 100 > 45): b[43] = 44
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * y) / 100 > 45): b[44] = 45
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * z) / 100 > 45): b[45] = 46
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 78 > 45): b[46] = 47
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 != 45): b[47] = 48
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 35): b[48] = 49
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 55): b[49] = 50
    if ((x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45) != (
            (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 110 > 45): b[50] = 51
    if (x ** 2 - y ** 2) / (z + 0.1) < -30 or (abs(x - y) * x) / 100 > 45:
        pattern_type = "111A2"
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((y + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35): b[51] = 52
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((z + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35): b[52] = 53
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + 42) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35): b[53] = 54
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + x) / 1.5) ** 2 < z * 20 and (x * y * z) / 1000 < 35): b[54] = 55
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + x) / 3) ** 2 < z * 20 and (x * y * z) / 1000 < 35): b[55] = 56
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + x) / 2) ** 1.8 < z * 20 and (x * y * z) / 1000 < 35): b[56] = 57
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + x) / 2) ** 3 < z * 20 and (x * y * z) / 1000 < 35): b[57] = 58
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + x) / 2) ** 2 != z * 20 and (x * y * z) / 1000 < 35): b[58] = 59
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + x) / 2) ** 2 < x * 20 and (x * y * z) / 1000 < 35): b[59] = 60
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + x) / 2) ** 2 < y * 20 and (x * y * z) / 1000 < 35): b[60] = 61
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + x) / 2) ** 2 < 52 * 20 and (x * y * z) / 1000 < 35): b[61] = 62
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + x) / 2) ** 2 < z * 12 and (x * y * z) / 1000 < 35): b[62] = 63
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + x) / 2) ** 2 < z * 20 or (x * y * z) / 1000 < 35): b[63] = 64
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + x) / 2) ** 2 < z * 20 and (y * y * z) / 1000 < 35): b[64] = 65
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + x) / 2) ** 2 < z * 20 and (z * y * z) / 1000 < 35): b[65] = 66
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + x) / 2) ** 2 < z * 20 and (x * x * z) / 1000 < 35): b[66] = 67
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + x) / 2) ** 2 < z * 20 and (x * z * z) / 1000 < 35): b[67] = 68
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + x) / 2) ** 2 < z * 20 and (x * y * x) / 1000 < 35): b[68] = 69
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + x) / 2) ** 2 < z * 20 and (x * y * y) / 1000 < 35): b[69] = 70
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + x) / 2) ** 2 < z * 20 and (52 * y * z) / 1000 < 35): b[70] = 71
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + x) / 2) ** 2 < z * 20 and (x * 38 * z) / 1000 < 35): b[71] = 72
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + x) / 2) ** 2 < z * 20 and (x * y * 41) / 1000 < 35): b[72] = 73
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 910 < 35): b[73] = 74
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 != 35): b[74] = 75
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 25): b[75] = 76
    if (((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35) != (
            ((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 44): b[76] = 77
    if ((x + x) / 2) ** 2 < z * 20 and (x * y * z) / 1000 < 35:
        pattern_type = "111A3"
    if (x < 30 and y > 45) != (x < 30 or y > 45): b[77] = 78
    if (x < 30 and y > 45) != (x != 30 and y > 45): b[78] = 79
    if (x < 30 and y > 45) != (x < 39 and y > 45): b[79] = 80
    if (x < 30 and y > 45) != (x < 30 and y != 45): b[80] = 81
    if (x < 30 and y > 45) != (x < 30 and y > 55): b[81] = 82
    if x < 30 and y > 45:
        pattern_type = 122
    if ((y > 60 and x < 50) or (y > 55 and x < 45 and z > 60)) != (
            (y != 60 and x < 50) or (y > 55 and x < 45 and z > 60)): b[82] = 83
    if ((y > 60 and x < 50) or (y > 55 and x < 45 and z > 60)) != (
            (y > 48 and x < 50) or (y > 55 and x < 45 and z > 60)): b[83] = 84
    if ((y > 60 and x < 50) or (y > 55 and x < 45 and z > 60)) != (
            (y > 60 or x < 50) or (y > 55 and x < 45 and z > 60)): b[84] = 85
    if ((y > 60 and x < 50) or (y > 55 and x < 45 and z > 60)) != (
            (y > 60 and x != 50) or (y > 55 and x < 45 and z > 60)): b[85] = 86
    if ((y > 60 and x < 50) or (y > 55 and x < 45 and z > 60)) != (
            (y > 60 and x < 58) or (y > 55 and x < 45 and z > 60)): b[86] = 87
    if ((y > 60 and x < 50) or (y > 55 and x < 45 and z > 60)) != (
            (y > 60 and x < 50) and (y > 55 and x < 45 and z > 60)): b[87] = 88
    if ((y > 60 and x < 50) or (y > 55 and x < 45 and z > 60)) != (
            (y > 60 and x < 50) or (y != 55 and x < 45 and z > 60)): b[88] = 89
    if ((y > 60 and x < 50) or (y > 55 and x < 45 and z > 60)) != (
            (y > 60 and x < 50) or (y > 45 and x < 45 and z > 60)): b[89] = 90
    if ((y > 60 and x < 50) or (y > 55 and x < 45 and z > 60)) != (
            (y > 60 and x < 50) or (y > 55 or x < 45 and z > 60)): b[90] = 91
    if ((y > 60 and x < 50) or (y > 55 and x < 45 and z > 60)) != (
            (y > 60 and x < 50) or (y > 55 and x != 45 and z > 60)): b[91] = 92
    if ((y > 60 and x < 50) or (y > 55 and x < 45 and z > 60)) != (
            (y > 60 and x < 50) or (y > 55 and x < 60 and z > 60)): b[92] = 93
    if ((y > 60 and x < 50) or (y > 55 and x < 45 and z > 60)) != (
            (y > 60 and x < 50) or (y > 55 and x < 45 or z > 60)): b[93] = 94
    if ((y > 60 and x < 50) or (y > 55 and x < 45 and z > 60)) != (
            (y > 60 and x < 50) or (y > 55 and x < 45 and z != 60)): b[94] = 95
    if ((y > 60 and x < 50) or (y > 55 and x < 45 and z > 60)) != (
            (y > 60 and x < 50) or (y > 55 and x < 45 and z > 69)): b[95] = 96
    if (y > 60 and x < 50) or (y > 55 and x < 45 and z > 60):
        pattern_type = 123
    if (z > 80 and x < 60) != (z > 80 or x < 60): b[96] = 97
    if (z > 80 and x < 60) != (z != 80 and x < 60): b[97] = 98
    if (z > 80 and x < 60) != (z > 89 and x < 60): b[98] = 99
    if (z > 80 and x < 60) != (z > 80 and x != 60): b[99] = 100
    if (z > 80 and x < 60) != (z > 80 and x < 46): b[100] = 101
    if z > 80 and x < 60:
        pattern_type = 124
    if ((30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 50 and y > 35)) != (
            (30 <= x < 47 and y > 30 and (x * y) / 100 < 18) or (x < 50 and y > 35)): b[101] = 102
    if ((30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 50 and y > 35)) != (
            (16 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 50 and y > 35)): b[102] = 103
    if ((30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 50 and y > 35)) != (
            (30 <= x < 55 or y > 30 and (x * y) / 100 < 18) or (x < 50 and y > 35)): b[103] = 104
    if ((30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 50 and y > 35)) != (
            (30 <= x < 55 and y != 30 and (x * y) / 100 < 18) or (x < 50 and y > 35)): b[104] = 105
    if ((30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 50 and y > 35)) != (
            (30 <= x < 55 and y > 41 and (x * y) / 100 < 18) or (x < 50 and y > 35)): b[105] = 106
    if ((30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 50 and y > 35)) != (
            (30 <= x < 55 and y > 30 or (x * y) / 100 < 18) or (x < 50 and y > 35)): b[106] = 107
    if ((30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 50 and y > 35)) != (
            (30 <= x < 55 and y > 30 and (25 * y) / 100 < 18) or (x < 50 and y > 35)): b[107] = 108
    if ((30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 50 and y > 35)) != (
            (30 <= x < 55 and y > 30 and (x * 45) / 100 < 18) or (x < 50 and y > 35)): b[108] = 109
    if ((30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 50 and y > 35)) != (
            (30 <= x < 55 and y > 30 and (y * y) / 100 < 18) or (x < 50 and y > 35)): b[109] = 110
    if ((30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 50 and y > 35)) != (
            (30 <= x < 55 and y > 30 and (x * x) / 100 < 18) or (x < 50 and y > 35)): b[110] = 111
    if ((30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 50 and y > 35)) != (
            (30 <= x < 55 and y > 30 and (x * y) / 89 < 18) or (x < 50 and y > 35)): b[111] = 112
    if ((30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 50 and y > 35)) != (
            (30 <= x < 55 and y > 30 and (x * y) / 100 < 25) or (x < 50 and y > 35)): b[112] = 113
    if ((30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 50 and y > 35)) != (
            (30 <= x < 55 and y > 30 and (x * y) / 100 != 18) or (x < 50 and y > 35)): b[113] = 114
    if ((30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 50 and y > 35)) != (
            (30 <= x < 55 and y > 30 and (x * y) / 100 < 18) and (x < 50 and y > 35)): b[114] = 115
    if ((30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 50 and y > 35)) != (
            (30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x != 50 and y > 35)): b[115] = 116
    if ((30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 50 and y > 35)) != (
            (30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 35 and y > 35)): b[116] = 117
    if ((30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 50 and y > 35)) != (
            (30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 50 or y > 35)): b[117] = 118
    if ((30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 50 and y > 35)) != (
            (30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 50 and y != 35)): b[118] = 119
    if ((30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 50 and y > 35)) != (
            (30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 50 and y > 45)): b[119] = 120
    if (30 <= x < 55 and y > 30 and (x * y) / 100 < 18) or (x < 50 and y > 35):
        pattern_type = 125
    if (((y ** 2) / (x + 1) > 30) or (y > 25 and x < 60)) != (((34 ** 2) / (x + 1) > 30) or (y > 25 and x < 60)): b[
        120] = 121
    if (((y ** 2) / (x + 1) > 30) or (y > 25 and x < 60)) != (((y ** 1.7) / (x + 1) > 30) or (y > 25 and x < 60)): b[
        121] = 122
    if (((y ** 2) / (x + 1) > 30) or (y > 25 and x < 60)) != (((y ** 2.4) / (x + 1) > 30) or (y > 25 and x < 60)): b[
        122] = 123
    if (((y ** 2) / (x + 1) > 30) or (y > 25 and x < 60)) != (((y ** 2) / (44 + 1) > 30) or (y > 25 and x < 60)): b[
        123] = 124
    if (((y ** 2) / (x + 1) > 30) or (y > 25 and x < 60)) != (((y ** 2) / (y + 1) > 30) or (y > 25 and x < 60)): b[
        124] = 125
    if (((y ** 2) / (x + 1) > 30) or (y > 25 and x < 60)) != (((y ** 2) / (x + 10) > 30) or (y > 25 and x < 60)): b[
        125] = 126
    if (((y ** 2) / (x + 1) > 30) or (y > 25 and x < 60)) != (((x ** 2) / (x + 1) > 30) or (y > 25 and x < 60)): b[
        126] = 127
    if (((y ** 2) / (x + 1) > 30) or (y > 25 and x < 60)) != (((y ** 2) / (x + 1) != 30) or (y > 25 and x < 60)): b[
        127] = 128
    if (((y ** 2) / (x + 1) > 30) or (y > 25 and x < 60)) != (((y ** 2) / (x + 1) > 44) or (y > 25 and x < 60)): b[
        128] = 129
    if (((y ** 2) / (x + 1) > 30) or (y > 25 and x < 60)) != (((y ** 2) / (x + 1) > 30) and (y > 25 and x < 60)): b[
        129] = 130
    if (((y ** 2) / (x + 1) > 30) or (y > 25 and x < 60)) != (((y ** 2) / (x + 1) > 30) or (y != 25 and x < 60)): b[
        130] = 131
    if (((y ** 2) / (x + 1) > 30) or (y > 25 and x < 60)) != (((y ** 2) / (x + 1) > 30) or (y > 14 and x < 60)): b[
        131] = 132
    if (((y ** 2) / (x + 1) > 30) or (y > 25 and x < 60)) != (((y ** 2) / (x + 1) > 30) or (y > 25 or x < 60)): b[
        132] = 133
    if (((y ** 2) / (x + 1) > 30) or (y > 25 and x < 60)) != (((y ** 2) / (x + 1) > 30) or (y > 25 and x != 60)): b[
        133] = 134
    if (((y ** 2) / (x + 1) > 30) or (y > 25 and x < 60)) != (((y ** 2) / (x + 1) > 30) or (y > 25 and x < 75)): b[
        134] = 135
    if (((y ** 2) / (x + 1) > 30) or (y > 25 and x < 60)) != (((y ** 2) / (x + 1) > 22) or (y > 25 and x < 60)): b[
        135] = 136
    if (((y ** 2) / (x + 1) > 30) or (y > 25 and x < 60)) != (((y ** 2) / (z + 1) > 30) or (y > 25 and x < 60)): b[
        136] = 137
    if (((y ** 2) / (x + 1) > 30) or (y > 25 and x < 60)) != (((z ** 2) / (x + 1) > 30) or (y > 25 and x < 60)): b[
        137] = 138
    if ((y ** 2) / (x + 1) > 30) or (y > 25 and x < 60):
        pattern_type = 126
    if (55 <= x < 75 and z > 40) != (55 <= x < 75 or z > 40): b[138] = 139
    if (55 <= x < 75 and z > 40) != (55 <= x < 75 and z != 40): b[139] = 140
    if (55 <= x < 75 and z > 40) != (55 <= x < 75 and z > 49): b[140] = 141
    if (55 <= x < 75 and z > 40) != (55 <= x < 85 and z > 40): b[141] = 142
    if (55 <= x < 75 and z > 40) != (39 <= x < 75 and z > 40): b[142] = 143
    if 55 <= x < 75 and z > 40:
        pattern_type = 127
    if ((15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)) != (
            (15 <= y <= 33 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)): b[143] = 144
    if ((15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)) != (
            (10 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)): b[144] = 145
    if ((15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)) != (
            (15 <= y <= 25 or x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)): b[145] = 146
    if ((15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)) != (
            (15 <= y <= 25 and x != 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)): b[146] = 147
    if ((15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)) != (
            (15 <= y <= 25 and x > 80 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)): b[147] = 148
    if ((15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)) != (
            (15 <= y <= 25 and x > 70 or (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)): b[148] = 149
    if ((15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)) != (
            (15 <= y <= 25 and x > 70 and (x * x) / 100 > 12) or (18 <= y <= 28 and x > 65)): b[149] = 150
    if ((15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)) != (
            (15 <= y <= 25 and x > 70 and (y * y) / 100 > 12) or (18 <= y <= 28 and x > 65)): b[150] = 151
    if ((15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)) != (
            (15 <= y <= 25 and x > 70 and (y * 24) / 100 > 12) or (18 <= y <= 28 and x > 65)): b[151] = 152
    if ((15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)) != (
            (15 <= y <= 25 and x > 70 and (33 * x) / 100 > 12) or (18 <= y <= 28 and x > 65)): b[152] = 153
    if ((15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)) != (
            (15 <= y <= 25 and x > 70 and (y * x) / 110 > 12) or (18 <= y <= 28 and x > 65)): b[153] = 154
    if ((15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)) != (
            (15 <= y <= 25 and x > 70 and (y * x) / 88 > 12) or (18 <= y <= 28 and x > 65)): b[154] = 155
    if ((15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)) != (
            (15 <= y <= 25 and x > 70 and (y * x) / 100 != 12) or (18 <= y <= 28 and x > 65)): b[155] = 156
    if ((15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)) != (
            (15 <= y <= 25 and x > 70 and (y * x) / 100 > 18) or (18 <= y <= 28 and x > 65)): b[156] = 157
    if ((15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)) != (
            (15 <= y <= 25 and x > 70 and (y * x) / 110 > 12) or (18 <= y <= 28 and x > 65)): b[157] = 158
    if ((15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)) != (
            (15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) and (18 <= y <= 28 and x > 65)): b[158] = 159
    if ((15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)) != (
            (15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (8 <= y <= 28 and x > 65)): b[159] = 160
    if ((15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)) != (
            (15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 38 and x > 65)): b[160] = 161
    if ((15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)) != (
            (15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 or x > 65)): b[161] = 162
    if ((15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)) != (
            (15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x != 65)): b[162] = 163
    if ((15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)) != (
            (15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 85)): b[163] = 164
    if ((15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65)) != (
            (15 <= y <= 25 and x > 70 and (y * x) / 100 > 5) or (18 <= y <= 28 and x > 65)): b[164] = 165
    if (15 <= y <= 25 and x > 70 and (y * x) / 100 > 12) or (18 <= y <= 28 and x > 65):
        pattern_type = 128
    if ((40 <= z < 60 and x > 65) or (45 <= z < 65 and x > 60)) != (
            (40 <= z < 69 and x > 65) or (45 <= z < 65 and x > 60)): b[165] = 166
    if ((40 <= z < 60 and x > 65) or (45 <= z < 65 and x > 60)) != (
            (20 <= z < 60 and x > 65) or (45 <= z < 65 and x > 60)): b[166] = 167
    if ((40 <= z < 60 and x > 65) or (45 <= z < 65 and x > 60)) != (
            (40 <= z < 60 or x > 65) or (45 <= z < 65 and x > 60)): b[167] = 168
    if ((40 <= z < 60 and x > 65) or (45 <= z < 65 and x > 60)) != (
            (40 <= z < 60 and x != 65) or (45 <= z < 65 and x > 60)): b[168] = 169
    if ((40 <= z < 60 and x > 65) or (45 <= z < 65 and x > 60)) != (
            (40 <= z < 60 and x > 49) or (45 <= z < 65 and x > 60)): b[169] = 170
    if ((40 <= z < 60 and x > 65) or (45 <= z < 65 and x > 60)) != (
            (40 <= z < 60 and x > 65) and (45 <= z < 65 and x > 60)): b[170] = 171
    if ((40 <= z < 60 and x > 65) or (45 <= z < 65 and x > 60)) != (
            (40 <= z < 60 and x > 65) or (25 <= z < 65 and x > 60)): b[171] = 172
    if ((40 <= z < 60 and x > 65) or (45 <= z < 65 and x > 60)) != (
            (40 <= z < 60 and x > 65) or (45 <= z < 85 and x > 60)): b[172] = 173
    if ((40 <= z < 60 and x > 65) or (45 <= z < 65 and x > 60)) != (
            (40 <= z < 60 and x > 65) or (45 <= z < 65 or x > 60)): b[173] = 174
    if ((40 <= z < 60 and x > 65) or (45 <= z < 65 and x > 60)) != (
            (40 <= z < 60 and x > 65) or (45 <= z < 65 and x != 60)): b[174] = 175
    if (40 <= z < 60 and x > 65) or (45 <= z < 65 and x > 60):
        pattern_type = 129
    if (x >= 85 and y <= 10 and z <= 30) != (x >= 75 and y <= 10 and z <= 30): b[175] = 176
    if (x >= 85 and y <= 10 and z <= 30) != (x >= 85 or y <= 10 and z <= 30): b[176] = 177
    if (x >= 85 and y <= 10 and z <= 30) != (x != 85 and y <= 10 and z <= 30): b[177] = 178
    if (x >= 85 and y <= 10 and z <= 30) != (x >= 85 and y != 10 and z <= 30): b[178] = 179
    if (x >= 85 and y <= 10 and z <= 30) != (x >= 85 and y <= 19 and z <= 30): b[179] = 180
    if (x >= 85 and y <= 10 and z <= 30) != (x >= 85 and y <= 10 or z <= 30): b[180] = 181
    if (x >= 85 and y <= 10 and z <= 30) != (x >= 85 and y <= 10 and z != 30): b[181] = 182
    if (x >= 85 and y <= 10 and z <= 30) != (x >= 85 and y <= 10 and z <= 39): b[182] = 183
    if x >= 85 and y <= 10 and z <= 30:
        pattern_type = 130

    # 返回被触发的规则编号集合
    return set(b.values())


def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """"""
    generated_path = execute_validation_rules(particle[0], particle[1], particle[2])

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # : x:1-50, y:1-50, z:1-50
        self.bounds = bounds if bounds else [(2, 100), (2, 100), (2, 100)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """target pathsPSO"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(particles[i][0], particles[i][1], particles[i][2]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(gbest_particle[0], gbest_particle[1], gbest_particle[2]),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """Path PSO"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - Path ")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations")
    print(f"Path : {len(target_paths)}")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"Path {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "" if result['success'] else f"({result['best_fitness']:.3f})"
        print(f"{status} | {result['time']:.2f}s | iterations{result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f": {success_count}/{len(target_paths)} ({success_rate:.1f}%) | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs}")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations, {len(target_paths)}Path ")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"---  {run_idx}/{num_runs} ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f": {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs} runcompleted | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 1: 
    ws1 = wb.active
    ws1.title = ""
    ws1.sheet_view.showGridLines = False

    headers = ["Run", "", "", "", "Average Iterations", "(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f" {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # 2: Path 
    ws2 = wb.create_sheet(title="Path ")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Path ID", "", "", "", "Average Iterations", "Minimum Iterations", "Maximum Iterations"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"Path  {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # 3: 
    ws3 = wb.create_sheet(title="")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Path ", "", "(x,y,z)", "", "Iterations", "Path "]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"{run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # 4: target paths
    ws4 = wb.create_sheet(title="target paths")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Path ID", "target paths", ""]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"Path  {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f" : {filename}")
    print(f"{'=' * 70}")
    print(f":")
    print(f"  1.        - {len(all_results)} run")
    print(f"  2. Path        - Path ")
    print(f"  3.    -  runPath ")
    print(f"  4. target paths       - target paths")
    print(f"{'=' * 70}\n")

    return filename


def main():


    target_paths = [
        {31, 36, 48, 53, 54, 55, 58, 60, 61, 63, 66, 67, 68, 71, 72, 76, 83, 85, 97, 98, 104, 105, 107, 118, 119, 121,
         123,
         127, 128, 131, 132, 133, 138, 139, 146, 162, 163, 168, 169, 174, 175},
        {2, 5, 6, 7, 8, 9, 10, 13, 17, 18, 23, 28, 30, 31, 34, 38, 39, 53, 56, 59, 60, 61, 62, 64, 78, 81, 83, 85, 97,
         98,
         107, 118, 119, 122, 124, 125, 126, 127, 130, 138, 146, 162, 163, 181},
        {31, 36, 48, 53, 54, 55, 58, 60, 61, 63, 66, 67, 68, 71, 72, 74, 76, 83, 85, 97, 98, 104, 105, 107, 118, 119,
         121,
         123, 128, 131, 132, 133, 138, 139, 146, 162, 163, 168, 169, 174, 175},
        {31, 36, 48, 53, 54, 55, 58, 60, 61, 63, 66, 67, 68, 71, 72, 74, 76, 78, 81, 83, 85, 89, 91, 94, 97, 98, 107,
         118,
         119, 121, 123, 128, 131, 132, 133, 138, 139, 146, 162, 163, 174, 175},
        {1, 3, 4, 11, 14, 28, 30, 31, 33, 34, 38, 39, 53, 56, 57, 59, 60, 61, 62, 64, 78, 81, 83, 85, 97, 98, 107, 118,
         119,
         122, 124, 125, 126, 127, 129, 130, 138, 146, 162, 163, 181},
        {11, 14, 31, 36, 48, 53, 58, 65, 66, 68, 71, 72, 78, 81, 83, 85, 89, 91, 94, 97, 98, 107, 118, 119, 122, 124,
         125,
         126, 127, 129, 130, 137, 139, 146, 162, 163, 174, 175},
        {31, 36, 48, 64, 65, 69, 70, 73, 75, 77, 83, 85, 89, 91, 94, 97, 98, 104, 105, 107, 118, 119, 121, 123, 127,
         128,
         131, 132, 133, 138, 139, 146, 162, 163, 174, 175},
        {1, 3, 4, 11, 12, 14, 28, 30, 31, 34, 38, 39, 53, 56, 57, 60, 61, 62, 64, 78, 81, 83, 85, 97, 98, 107, 118, 119,
         122, 124, 125, 127, 130, 138, 146, 162, 163, 181},
        {31, 36, 48, 53, 54, 55, 58, 60, 61, 66, 68, 71, 72, 78, 81, 83, 85, 97, 98, 107, 118, 119, 121, 123, 128, 131,
         132,
         133, 136, 138, 139, 146, 162, 163, 168, 169},
        {14, 24, 39, 40, 42, 43, 44, 45, 46, 50, 51, 78, 79, 94, 97, 100, 116, 118, 123, 124, 125, 127, 128, 133, 134,
         136,
         138, 139, 146, 149, 162, 168, 173, 174, 177},
        {14, 22, 24, 28, 30, 31, 33, 34, 35, 38, 39, 59, 61, 62, 64, 78, 81, 83, 85, 97, 98, 107, 118, 119, 121, 123,
         128,
         131, 132, 133, 136, 137, 146, 162, 163, 181},
        {11, 14, 31, 36, 48, 53, 58, 65, 66, 68, 71, 72, 76, 78, 81, 83, 85, 89, 91, 94, 99, 107, 118, 119, 122, 124,
         125,
         126, 127, 129, 130, 137, 139, 146, 162, 163},
        {2, 5, 6, 7, 8, 9, 10, 13, 16, 17, 18, 20, 22, 23, 26, 28, 30, 31, 32, 34, 38, 39, 56, 57, 59, 61, 62, 64, 78,
         81,
         83, 85, 97, 98, 103, 104, 107, 118, 181},
        {14, 22, 24, 27, 29, 32, 36, 37, 48, 59, 60, 61, 62, 64, 78, 81, 83, 85, 97, 98, 107, 118, 119, 121, 123, 128,
         131,
         132, 133, 136, 137, 146, 162, 163, 181},
        {14, 24, 27, 29, 30, 36, 41, 44, 47, 48, 49, 59, 64, 78, 79, 116, 118, 123, 124, 125, 127, 128, 133, 134, 136,
         137,
         146, 149, 162, 168, 174, 177, 179, 181},
        {14, 23, 24, 27, 29, 30, 36, 37, 40, 41, 43, 44, 48, 78, 79, 85, 86, 91, 116, 118, 121, 122, 130, 138, 146, 149,
         162, 167, 168, 172, 174, 177, 179, 181},
        {3, 11, 14, 27, 32, 36, 37, 48, 52, 54, 55, 58, 63, 78, 81, 83, 85, 97, 98, 107, 118, 119, 122, 124, 125, 126,
         127,
         129, 130, 138, 146, 162, 163, 181},
        {31, 36, 48, 53, 54, 55, 58, 60, 61, 62, 63, 66, 67, 68, 72, 83, 85, 89, 91, 94, 97, 98, 104, 105, 107, 118,
         119,
         127, 128, 131, 133, 138, 139, 143},
        {14, 24, 27, 29, 30, 36, 44, 48, 59, 64, 78, 79, 116, 118, 123, 124, 125, 127, 128, 133, 134, 135, 136, 137,
         139,
         140, 146, 149, 162, 168, 174, 181},
        {14, 24, 39, 40, 42, 43, 44, 45, 46, 50, 51, 78, 79, 94, 116, 118, 123, 124, 125, 127, 128, 133, 134, 136, 137,
         138,
         139, 146, 149, 162, 171, 177},
        {14, 23, 24, 27, 30, 31, 33, 34, 36, 37, 38, 48, 78, 79, 85, 86, 91, 92, 94, 97, 100, 116, 118, 121, 122, 130,
         139,
         142, 146, 149, 162, 168, 174},
        {14, 24, 39, 40, 42, 43, 45, 46, 50, 51, 116, 118, 123, 124, 125, 127, 128, 133, 134, 137, 146, 149, 161, 162,
         167,
         168, 172, 174, 177, 179, 181},
        {14, 24, 39, 40, 42, 43, 44, 45, 46, 50, 94, 116, 118, 123, 124, 125, 127, 128, 133, 134, 138, 139, 146, 149,
         161,
         162, 166, 168, 173, 174, 177},
        {14, 24, 27, 30, 36, 37, 38, 48, 78, 79, 85, 86, 87, 91, 97, 98, 104, 107, 108, 114, 116, 118, 139, 143, 149,
         168,
         169, 170, 174, 175},
        {11, 14, 27, 31, 33, 34, 36, 37, 38, 48, 53, 58, 65, 66, 68, 71, 74, 76, 78, 81, 83, 85, 97, 98, 115, 120, 139,
         168,
         169, 174, 175},
        {4, 5, 6, 8, 10, 17, 18, 20, 23, 28, 31, 32, 39, 52, 53, 54, 55, 58, 60, 63, 65, 66, 70, 71, 73, 74, 76, 88, 97,
         98,
         115, 149, 181},
        {39, 40, 41, 42, 43, 44, 45, 46, 50, 51, 94, 107, 127, 128, 138, 139, 142, 148, 151, 152, 154, 157, 158, 159,
         166,
         168, 173, 174},
        {39, 40, 41, 42, 43, 44, 45, 50, 51, 94, 97, 100, 123, 127, 128, 133, 134, 138, 139, 144, 146, 149, 161, 162,
         168,
         173, 174, 177},
        {2, 5, 6, 7, 8, 9, 10, 13, 17, 18, 20, 22, 23, 26, 28, 30, 31, 32, 39, 56, 57, 59, 61, 62, 64, 82, 83, 85, 97,
         98,
         115, 149, 181},
        {4, 5, 6, 8, 10, 13, 17, 18, 20, 21, 22, 23, 26, 28, 31, 32, 39, 78, 79, 80, 88, 97, 98, 115, 117, 139, 149,
         168,
         169, 174, 175},
        {4, 5, 6, 10, 17, 18, 20, 21, 22, 23, 26, 28, 32, 39, 78, 79, 85, 87, 91, 92, 93, 94, 97, 98, 104, 107, 114,
         118,
         139, 143, 149},
        {2, 5, 6, 9, 10, 16, 17, 18, 20, 21, 22, 23, 26, 28, 30, 31, 32, 39, 56, 59, 61, 62, 64, 82, 83, 84, 85, 97, 98,
         115, 149, 181},
        {39, 40, 41, 42, 44, 45, 50, 51, 94, 107, 127, 128, 138, 139, 142, 146, 149, 150, 153, 155, 156, 160, 162, 165,
         168,
         173, 174},
        {36, 48, 97, 98, 104, 105, 107, 123, 127, 128, 131, 132, 133, 138, 139, 143, 146, 147, 149, 162, 163, 168, 169,
         170,
         174, 175},
        {5, 6, 7, 10, 17, 18, 20, 22, 23, 26, 28, 31, 32, 39, 56, 57, 59, 61, 62, 64, 83, 84, 85, 91, 94, 95, 97, 98,
         115,
         149, 181},
        {14, 15, 16, 19, 24, 25, 28, 31, 32, 39, 78, 79, 85, 86, 87, 91, 92, 93, 94, 97, 98, 104, 107, 114, 116, 118,
         139,
         143, 149},
        {27, 30, 31, 33, 36, 37, 38, 48, 78, 79, 85, 86, 87, 91, 92, 93, 94, 99, 101, 104, 107, 108, 114, 116, 118, 139,
         143, 149},
        {32, 36, 48, 52, 53, 56, 57, 59, 60, 61, 62, 64, 78, 81, 83, 85, 97, 98, 107, 118, 119, 121, 128, 131, 133, 177,
         178, 181},
        {11, 14, 27, 31, 33, 34, 36, 37, 38, 48, 64, 67, 69, 70, 72, 73, 75, 77, 82, 83, 84, 85, 89, 90, 91, 94, 99,
         115,
         139},
        {14, 24, 36, 48, 116, 118, 123, 124, 125, 127, 128, 133, 134, 135, 136, 137, 138, 141, 149, 168, 169, 170, 172,
         174},
        {14, 24, 39, 40, 42, 43, 45, 46, 107, 127, 128, 145, 146, 149, 160, 162, 167, 168, 172, 174, 177, 179, 180,
         181},
        {39, 40, 41, 42, 43, 44, 45, 46, 50, 59, 64, 107, 127, 128, 160, 162, 167, 168, 172, 174, 176, 177, 178, 181},
        {39, 40, 41, 42, 43, 44, 45, 46, 50, 51, 94, 123, 127, 128, 133, 134, 138, 139, 142, 159, 164, 168, 173, 174},
        {1, 2, 9, 11, 14, 28, 29, 30, 32, 39, 64, 67, 69, 72, 73, 75, 77, 88, 96, 97, 98, 115, 139, 174, 175},
        {39, 40, 42, 43, 44, 45, 46, 59, 64, 107, 127, 128, 160, 162, 167, 168, 172, 174, 177, 181, 182, 183},
        {36, 48, 97, 98, 104, 107, 108, 110, 113, 114, 116, 118, 130, 139, 143, 149, 168, 169, 170, 174, 175},
        {36, 48, 83, 85, 97, 98, 102, 106, 109, 111, 112, 115, 130, 139, 143, 149, 168, 169, 174, 175},
    ]

    print("=" * 70)
    print("baseline PSO")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("Program completed")


if __name__ == "__main__":
    main()