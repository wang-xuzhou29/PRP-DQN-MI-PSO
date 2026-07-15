import os
import random
from collections import deque
import numpy as np
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
STATE_MIN_X, STATE_MAX_X = 1, 30    # cpu_cores 范围 1~30
STATE_MIN_Y, STATE_MAX_Y = 1, 40    # memory_gb 范围 1~40
STATE_MIN_Z, STATE_MAX_Z = 1, 2100  # disk_space_gb 范围 1~2100
def generate_input():
    return [
        random.randint(STATE_MIN_X, STATE_MAX_X),
        random.randint(STATE_MIN_Y, STATE_MAX_Y),
        random.randint(STATE_MIN_Z, STATE_MAX_Z)
    ]


import os
import random
from collections import deque
import numpy as np
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# --- 全局状态范围配置 ---
STATE_MIN_X, STATE_MAX_X = 1, 100
STATE_MIN_Y, STATE_MAX_Y = 1, 100
STATE_MIN_Z, STATE_MAX_Z = 1, 100
def generate_input():
    return [
        random.randint(STATE_MIN_X, STATE_MAX_X),
        random.randint(STATE_MIN_Y, STATE_MAX_Y),
        random.randint(STATE_MIN_Z, STATE_MAX_Z)
    ]

# === 执行规则函数 (更新后) ===
def execute_Tr(a):
    """
    替换原有的execute_Tr函数
    参数a: 包含3个元素的元组或数组，分别对应cpu_cores, memory_gb, disk_space_gb
    返回: 触发的规则编号集合
    """
    cpu_cores, memory_gb, disk_space_gb = int(a[0]), float(a[1]), float(a[2])
    triggered = set()

    # 创建一个字典来存储b数组的值，用于跟踪哪些规则被触发
    b = {}

    # 按照原始代码的逻辑进行条件判断
    if (cpu_cores >= 16) != (cpu_cores >= 13):
        b[0] = 1
        triggered.add(1)
    if (cpu_cores >= 16) != (cpu_cores >= 18):
        b[1] = 2
        triggered.add(2)
    if (cpu_cores >= 16) != (cpu_cores >= 21):
        b[2] = 3
        triggered.add(3)
    if (cpu_cores >= 16) != (cpu_cores >= 24):
        b[3] = 4
        triggered.add(4)

    if cpu_cores >= 16:
        if (memory_gb >= 32.0 and disk_space_gb >= 1000) != (memory_gb != 32.0 and disk_space_gb >= 1000):
            b[4] = 5
            triggered.add(5)
        if (memory_gb >= 32.0 and disk_space_gb >= 1000) != (memory_gb == 32.0 and disk_space_gb >= 1000):
            b[5] = 6
            triggered.add(6)
        if (memory_gb >= 32.0 and disk_space_gb >= 1000) != (memory_gb >= 32.0 and disk_space_gb == 1000):
            b[6] = 7
            triggered.add(7)
        if (memory_gb >= 32.0 and disk_space_gb >= 1000) != (memory_gb >= 32.0 and disk_space_gb != 1000):
            b[7] = 8
            triggered.add(8)
        if (memory_gb >= 32.0 and disk_space_gb >= 1000) != (memory_gb >= 38.0 and disk_space_gb >= 1000):
            b[8] = 9
            triggered.add(9)
        if (memory_gb >= 32.0 and disk_space_gb >= 1000) != (memory_gb >= 32.0 and disk_space_gb >= 1750):
            b[9] = 10
            triggered.add(10)
        if (memory_gb >= 32.0 and disk_space_gb >= 1000) != (memory_gb >= 45.0 and disk_space_gb >= 1000):
            b[10] = 11
            triggered.add(11)

    if memory_gb >= 32.0 and disk_space_gb >= 1000:
        if (disk_space_gb >= 2000) != (disk_space_gb != 2000):
            b[11] = 12
            triggered.add(12)
        if (disk_space_gb >= 2000) != (disk_space_gb >= 1540):
            b[12] = 13
            triggered.add(13)
        if (disk_space_gb >= 2000) != (disk_space_gb >= 1670):
            b[13] = 14
            triggered.add(14)
        if (disk_space_gb >= 1500) != (disk_space_gb != 1500):
            b[14] = 15
            triggered.add(15)
        if (disk_space_gb >= 1500) != (disk_space_gb == 1500):
            b[15] = 16
            triggered.add(16)
        if (disk_space_gb >= 1500) != (disk_space_gb >= 1770):
            b[16] = 17
            triggered.add(17)

    if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb != 16.0 and disk_space_gb >= 500):
        b[17] = 18
        triggered.add(18)
    if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb == 16.0 and disk_space_gb >= 500):
        b[18] = 19
        triggered.add(19)
    if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 26.0 and disk_space_gb >= 500):
        b[19] = 20
        triggered.add(20)
    if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb != 500):
        b[20] = 21
        triggered.add(21)
    if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb == 500):
        b[21] = 22
        triggered.add(22)
    if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb >= 900):
        b[22] = 23
        triggered.add(23)
    if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 33.0 and disk_space_gb >= 500):
        b[23] = 24
        triggered.add(24)
    if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb >= 780):
        b[24] = 25
        triggered.add(25)

    elif memory_gb >= 16.0 and disk_space_gb >= 500:
        if (disk_space_gb >= 1000) != (disk_space_gb != 1000):
            b[25] = 26
            triggered.add(26)
        if (disk_space_gb >= 1000) != (disk_space_gb == 1000):
            b[26] = 27
            triggered.add(27)
        if (disk_space_gb >= 1000) != (disk_space_gb >= 1200):
            b[27] = 28
            triggered.add(28)

    if (memory_gb >= 8.0) != (memory_gb != 8.0):
        b[28] = 29
        triggered.add(29)
    if (memory_gb >= 8.0) != (memory_gb == 8.0):
        b[29] = 30
        triggered.add(30)
    if (memory_gb >= 8.0) != (memory_gb >= 12.0):
        b[30] = 31
        triggered.add(31)

    if (cpu_cores >= 8) != (cpu_cores != 8):
        b[31] = 32
        triggered.add(32)
    if (cpu_cores >= 8) != (cpu_cores == 8):
        b[32] = 33
        triggered.add(33)
    if (cpu_cores >= 8) != (cpu_cores >= 9):
        b[33] = 34
        triggered.add(34)

    elif cpu_cores >= 8:
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb != 16.0 and disk_space_gb >= 500):
            b[34] = 35
            triggered.add(35)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb == 16.0 and disk_space_gb >= 500):
            b[35] = 36
            triggered.add(36)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb != 500):
            b[36] = 37
            triggered.add(37)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb == 500):
            b[37] = 38
            triggered.add(38)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 22.0 and disk_space_gb >= 500):
            b[38] = 39
            triggered.add(39)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb >= 800):
            b[39] = 40
            triggered.add(40)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 27.0 and disk_space_gb >= 500):
            b[40] = 41
            triggered.add(41)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb >= 940):
            b[41] = 42
            triggered.add(42)

    if memory_gb >= 16.0 and disk_space_gb >= 500:
        if (disk_space_gb >= 1000 and memory_gb >= 32.0) != (disk_space_gb != 1000 and memory_gb >= 32.0):
            b[42] = 43
            triggered.add(43)
        if (disk_space_gb >= 1000 and memory_gb >= 32.0) != (disk_space_gb == 1000 and memory_gb >= 32.0):
            b[43] = 44
            triggered.add(44)
        if (disk_space_gb >= 1000 and memory_gb >= 32.0) != (disk_space_gb >= 1000 and memory_gb != 32.0):
            b[44] = 45
            triggered.add(45)
        if (disk_space_gb >= 1000 and memory_gb >= 32.0) != (disk_space_gb >= 1000 and memory_gb == 32.0):
            b[45] = 46
            triggered.add(46)
        if (disk_space_gb >= 1000 and memory_gb >= 32.0) != (disk_space_gb >= 1000 and memory_gb >= 36.0):
            b[46] = 47
            triggered.add(47)
        if (disk_space_gb >= 1000 and memory_gb >= 32.0) != (disk_space_gb >= 1100 and memory_gb >= 32.0):
            b[47] = 48
            triggered.add(48)

    if (disk_space_gb >= 750) != (disk_space_gb != 750):
        b[48] = 49
        triggered.add(49)
    if (disk_space_gb >= 750) != (disk_space_gb == 750):
        b[49] = 50
        triggered.add(50)
    if (disk_space_gb >= 750) != (disk_space_gb >= 850):
        b[50] = 51
        triggered.add(51)

    if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb != 8.0 and disk_space_gb >= 250):
        b[51] = 52
        triggered.add(52)
    if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb == 8.0 and disk_space_gb >= 250):
        b[52] = 53
        triggered.add(53)
    if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb >= 8.0 and disk_space_gb != 250):
        b[53] = 54
        triggered.add(54)
    if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb >= 8.0 and disk_space_gb == 250):
        b[54] = 55
        triggered.add(55)
    if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb >= 8.0 and disk_space_gb >= 280):
        b[55] = 56
        triggered.add(56)
    if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb >= 10.0 and disk_space_gb >= 250):
        b[56] = 57
        triggered.add(57)

    if (cpu_cores >= 4) != (cpu_cores != 4):
        b[57] = 58
        triggered.add(58)
    if (cpu_cores >= 4) != (cpu_cores >= 6):
        b[58] = 59
        triggered.add(59)

    elif cpu_cores >= 4:
        if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb != 9.5 and disk_space_gb >= 250):
            b[59] = 60
            triggered.add(60)
        if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb == 8.0 and disk_space_gb >= 250):
            b[60] = 61
            triggered.add(61)
        if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb >= 8.0 and disk_space_gb != 250):
            b[61] = 62
            triggered.add(62)
        if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb >= 8.0 and disk_space_gb == 250):
            b[62] = 63
            triggered.add(63)

    if memory_gb >= 8.0 and disk_space_gb >= 250:
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb != 16.0 and disk_space_gb >= 500):
            b[63] = 64
            triggered.add(64)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb == 16.0 and disk_space_gb >= 500):
            b[64] = 65
            triggered.add(65)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb != 500):
            b[65] = 66
            triggered.add(66)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb == 500):
            b[66] = 67
            triggered.add(67)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 19.0 and disk_space_gb >= 500):
            b[67] = 68
            triggered.add(68)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb >= 660):
            b[68] = 69
            triggered.add(69)

    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb != 4.0 and disk_space_gb >= 100):
        b[69] = 70
        triggered.add(70)
    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb == 4.0 and disk_space_gb >= 100):
        b[70] = 71
        triggered.add(71)
    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb >= 4.0 and disk_space_gb != 100):
        b[71] = 72
        triggered.add(72)
    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb >= 6.2 and disk_space_gb >= 100):
        b[72] = 73
        triggered.add(73)
    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb >= 7.0 and disk_space_gb >= 210):
        b[73] = 74
        triggered.add(74)

    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb != 4.0 and disk_space_gb >= 100):
        b[74] = 75
        triggered.add(75)
    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb >= 5.3 and disk_space_gb >= 100):
        b[75] = 76
        triggered.add(76)
    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb >= 4.0 and disk_space_gb != 100):
        b[76] = 77
        triggered.add(77)
    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb >= 4.0 and disk_space_gb >= 278):
        b[77] = 78
        triggered.add(78)

    # 内存分配计算基于三维资源
    if (memory_gb >= 32.0) != (memory_gb == 32.0):
        b[78] = 79
        triggered.add(79)
    if (memory_gb >= 32.0) != (memory_gb >= 36.5):
        b[79] = 80
        triggered.add(80)
    if (memory_gb >= 32.0) != (memory_gb >= 35.0):
        b[80] = 81
        triggered.add(81)

    if memory_gb >= 32.0:
        if (cpu_cores >= 16 and disk_space_gb >= 1000) != (cpu_cores >= 16 and disk_space_gb != 1000):
            b[81] = 82
            triggered.add(82)

    if (cpu_cores >= 16 and disk_space_gb >= 1000) != (cpu_cores >= 16 and disk_space_gb == 1000):
        b[82] = 83
        triggered.add(83)
    if (cpu_cores >= 16 and disk_space_gb >= 1000) != (cpu_cores >= 22 and disk_space_gb >= 1000):
        b[83] = 84
        triggered.add(84)
    if (cpu_cores >= 16 and disk_space_gb >= 1000) != (cpu_cores == 16 and disk_space_gb >= 1000):
        b[84] = 85
        triggered.add(85)
    if (cpu_cores >= 16 and disk_space_gb >= 1000) != (cpu_cores != 16 and disk_space_gb >= 1000):
        b[85] = 86
        triggered.add(86)
    if (cpu_cores >= 16 and disk_space_gb >= 1000) != (cpu_cores >= 31 and disk_space_gb >= 1000):
        b[86] = 87
        triggered.add(87)

    if (cpu_cores >= 8) != (cpu_cores != 8):
        b[87] = 88
        triggered.add(88)
    if (cpu_cores >= 8) != (cpu_cores >= 9.4):
        b[88] = 89
        triggered.add(89)

    if (memory_gb >= 16.0) != (memory_gb >= 22):
        b[89] = 90
        triggered.add(90)
    if (memory_gb >= 16.0) != (memory_gb >= 31):
        b[90] = 91
        triggered.add(91)

    elif memory_gb >= 16.0:
        if (cpu_cores >= 8 and disk_space_gb >= 500) != (cpu_cores >= 8 and disk_space_gb >= 950):
            b[91] = 92
            triggered.add(92)
        if (cpu_cores >= 8 and disk_space_gb >= 500) != (cpu_cores == 8 and disk_space_gb >= 500):
            b[92] = 93
            triggered.add(93)
        if (cpu_cores >= 8 and disk_space_gb >= 500) != (cpu_cores >= 8 and disk_space_gb == 500):
            b[93] = 94
            triggered.add(94)
        if (cpu_cores >= 8 and disk_space_gb >= 500) != (cpu_cores >= 8 and disk_space_gb != 500):
            b[94] = 95
            triggered.add(95)
        if (cpu_cores >= 8 and disk_space_gb >= 500) != (cpu_cores != 8 and disk_space_gb >= 500):
            b[95] = 96
            triggered.add(96)
        if (cpu_cores >= 8 and disk_space_gb >= 500) != (cpu_cores >= 8 or disk_space_gb >= 500):
            b[96] = 97
            triggered.add(97)

    if (memory_gb >= 8.0) != (memory_gb != 8.0):
        b[97] = 98
        triggered.add(98)
    if (memory_gb >= 8.0) != (memory_gb >= 11):
        b[98] = 99
        triggered.add(99)
    if (memory_gb >= 8.0) != (memory_gb >= 12.0):
        b[99] = 100
        triggered.add(100)

    return triggered



def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0

# === 目标路径组 ===
targetPaths = [
    # A1
    {2, 3, 4, 6, 7, 9, 10, 11, 12, 13, 14, 16, 17, 19, 22, 27, 30, 33, 36, 38, 44, 46, 47, 50, 53, 55, 61, 63, 65, 67,
     71, 79, 80, 81, 83, 84, 85, 87, 93, 94},
    # A2
    {3, 4, 6, 7, 9, 10, 11, 12, 13, 14, 16, 17, 19, 22, 27, 30, 33, 36, 38, 44, 46, 47, 50, 53, 55, 61, 63, 65, 67, 71,
     79, 80, 81, 83, 84, 85, 87, 93, 94},
    # A3
    {2, 3, 4, 6, 7, 9, 10, 11, 12, 15, 19, 22, 27, 28, 30, 33, 36, 38, 44, 46, 47, 48, 50, 53, 55, 61, 63, 65, 67, 71,
     79, 80, 81, 83, 84, 86, 87, 93, 94},
    # A4
    {3, 4, 5, 7, 9, 10, 11, 12, 13, 16, 17, 19, 22, 24, 27, 30, 33, 36, 38, 44, 45, 47, 50, 53, 55, 61, 63, 65, 67, 71,
     80, 81, 83, 84, 85, 87, 93, 94},
    # A5
    {2, 3, 4, 5, 19, 20, 22, 24, 27, 28, 30, 33, 36, 38, 39, 41, 45, 50, 53, 55, 61, 63, 65, 67, 68, 71, 83, 84, 85, 87,
     90, 91},
    # A6
    {3, 4, 5, 18, 20, 22, 24, 27, 28, 30, 33, 35, 38, 39, 41, 45, 50, 53, 55, 61, 63, 64, 67, 68, 71, 83, 84, 85, 87,
     90, 91},
    # A7
    {3, 4, 8, 19, 22, 23, 25, 30, 33, 36, 38, 40, 42, 43, 49, 53, 55, 61, 63, 65, 67, 69, 71, 79, 80, 81, 82, 92, 93,
     94},
    # A8
    {3, 4, 8, 19, 22, 23, 25, 30, 33, 36, 38, 40, 42, 43, 50, 51, 53, 55, 61, 63, 65, 67, 71, 79, 80, 81, 82, 92, 93,
     94},
    # A9
    {12, 13, 14, 16, 19, 22, 24, 27, 30, 32, 34, 44, 45, 47, 50, 53, 55, 61, 63, 65, 67, 71, 80, 81, 86, 88, 89, 94,
     96},
    # A10
    {1, 12, 15, 19, 22, 24, 27, 28, 30, 33, 36, 38, 44, 45, 47, 48, 50, 53, 55, 61, 63, 65, 67, 71, 80, 81, 86, 93, 94},
    # A11
    {2, 3, 4, 18, 20, 22, 23, 24, 26, 30, 33, 35, 38, 39, 40, 41, 42, 50, 51, 53, 55, 61, 63, 64, 67, 68, 71, 90, 91},
    # A12
    {12, 13, 14, 16, 17, 19, 22, 24, 27, 30, 32, 44, 45, 47, 50, 53, 55, 61, 63, 65, 67, 71, 80, 81, 86, 88, 96, 97},
    # A13
    {12, 13, 14, 16, 17, 19, 22, 24, 27, 30, 32, 44, 45, 47, 50, 53, 55, 58, 65, 67, 71, 80, 81, 86, 88, 96, 97},
    # A14
    {12, 13, 14, 16, 17, 19, 22, 27, 30, 32, 44, 46, 47, 50, 53, 55, 59, 65, 67, 71, 79, 80, 81, 86, 88, 96, 97},
    # A15
    {3, 4, 19, 21, 23, 24, 25, 30, 33, 36, 37, 40, 42, 49, 53, 55, 61, 63, 65, 66, 69, 71, 91},
    # A16
    {2, 3, 4, 8, 21, 30, 33, 37, 49, 53, 55, 56, 61, 63, 66, 71, 78, 79, 80, 81, 82, 95, 97},
    # A17
    {2, 3, 4, 5, 18, 29, 33, 35, 50, 52, 60, 70, 73, 74, 75, 76, 83, 84, 85, 87, 98},
    # A18
    {2, 3, 4, 8, 21, 30, 33, 37, 49, 54, 62, 71, 74, 78, 79, 80, 81, 82, 95, 97},
    # A19
    {2, 3, 4, 18, 30, 31, 33, 35, 50, 51, 53, 55, 57, 61, 63, 64, 71, 99, 100},
    # A20
    {2, 3, 4, 8, 21, 30, 33, 37, 49, 54, 62, 72, 77, 79, 80, 81, 82, 95, 97}
]


class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 2000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(targetPaths)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }


def compute_robustness(state, path):
    base = execute_Tr(*state)
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                        np.clip(state[0] + dx, STATE_MIN_X, STATE_MAX_X),
                        np.clip(state[1] + dy, STATE_MIN_Y, STATE_MAX_Y),
                        np.clip(state[2] + dz, STATE_MIN_Z, STATE_MAX_Z)
                    ])
                n_trig = execute_Tr(*neighbor)
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = targetPaths[target_path_idx]
    samples = []
    attempts = 0

    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(STATE_MIN_X, STATE_MAX_X),
            random.randint(STATE_MIN_Y, STATE_MAX_Y),
            random.randint(STATE_MIN_Z, STATE_MAX_Z)
        ])
        triggered = execute_Tr(*state)

        if not triggered:
            continue

        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)

        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }

        samples.append(sample_data)

    return samples


def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample[
                'robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)

        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]

    return selected_samples


def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = targetPaths[target_path_idx]
        samples = []
        attempts = 0

        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(STATE_MIN_X, STATE_MAX_X),
                random.randint(STATE_MIN_Y, STATE_MAX_Y),
                random.randint(STATE_MIN_Z, STATE_MAX_Z)
            ])
            triggered = execute_Tr(*state)

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)

            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)

        random.shuffle(samples)
        return samples[:config.top_k_samples]

    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")

        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)


def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}

    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates

    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}

        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config,
                                                         shared_candidates[path_idx])

            strategy_results[path_idx] = samples

        results[strategy_name] = strategy_results

    return results


def analyze_fitness_values(results, config):
    analysis_results = {}

    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []

        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])

        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }

        all_scores = []

        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])

        scores_array = np.array(all_scores)

        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)

        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })

        analysis_results[strategy_name] = analysis

    return analysis_results


def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []

    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)

    df = pd.DataFrame(df_data)
    return df, analysis_results


def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []

    print(f"Starting {num_runs} experiments...")

    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")

        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)

        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)

        print(f"Experiment {run_idx + 1} completed.")

    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df


def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')

    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width

    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []

    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]

        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)

    stats_df = pd.DataFrame(stats_data)

    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Results saved to: {output_path}")


def main():
    results_df = run_multiple_experiments(num_runs=20)

    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")

    save_results_to_excel(results_df, output_path)

    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)

    return results_df, output_path


if __name__ == "__main__":
    results, output_file = main()

def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if set2.issubset(set1):
        return 1.0

    return intersection / union if union != 0 else 0.0


targetPaths = [
    {1, 2, 3, 4, 10, 11, 12, 13, 14, 15, 24, 25, 26, 27, 28, 29},
    {5, 6, 7, 8, 9, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25},
    {5, 6, 7, 8, 9, 17, 18, 19, 20, 21, 24, 25, 26, 27, 28, 29}
]


class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 2000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(targetPaths)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }


def compute_robustness(state, path):
    base = execute_Tr(*state)
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                        np.clip(state[0] + dx, STATE_MIN_X, STATE_MAX_X),
                        np.clip(state[1] + dy, STATE_MIN_Y, STATE_MAX_Y),
                        np.clip(state[2] + dz, STATE_MIN_Z, STATE_MAX_Z)
                    ])
                n_trig = execute_Tr(*neighbor)
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = targetPaths[target_path_idx]
    samples = []
    attempts = 0

    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(STATE_MIN_X, STATE_MAX_X),
            random.randint(STATE_MIN_Y, STATE_MAX_Y),
            random.randint(STATE_MIN_Z, STATE_MAX_Z)
        ])
        triggered = execute_Tr(*state)

        if not triggered:
            continue

        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)

        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }

        samples.append(sample_data)

    return samples


def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample[
                'robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)

        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]

    return selected_samples


def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = targetPaths[target_path_idx]
        samples = []
        attempts = 0

        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(STATE_MIN_X, STATE_MAX_X),
                random.randint(STATE_MIN_Y, STATE_MAX_Y),
                random.randint(STATE_MIN_Z, STATE_MAX_Z)
            ])
            triggered = execute_Tr(*state)

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)

            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)

        random.shuffle(samples)
        return samples[:config.top_k_samples]

    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")

        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)


def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}

    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates

    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}

        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config,
                                                         shared_candidates[path_idx])

            strategy_results[path_idx] = samples

        results[strategy_name] = strategy_results

    return results


def analyze_fitness_values(results, config):
    analysis_results = {}

    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []

        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])

        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }

        all_scores = []

        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])

        scores_array = np.array(all_scores)

        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)

        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })

        analysis_results[strategy_name] = analysis

    return analysis_results


def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []

    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)

    df = pd.DataFrame(df_data)
    return df, analysis_results


def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []

    print(f"Starting {num_runs} experiments...")

    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")

        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)

        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)

        print(f"Experiment {run_idx + 1} completed.")

    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df


def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')

    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width

    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []

    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]

        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)

    stats_df = pd.DataFrame(stats_data)

    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Results saved to: {output_path}")


def main():
    results_df = run_multiple_experiments(num_runs=20)

    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")

    save_results_to_excel(results_df, output_path)

    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)

    return results_df, output_path


if __name__ == "__main__":
    results, output_file = main()