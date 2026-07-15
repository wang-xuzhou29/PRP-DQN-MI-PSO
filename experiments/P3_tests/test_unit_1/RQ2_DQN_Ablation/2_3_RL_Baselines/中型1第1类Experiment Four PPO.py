import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
import numpy as np
import random
import time
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
from datetime import datetime

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# ==================== 状态范围（已修改） ====================
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'ACTION_DIM': 3,
    'MIN_VALUES': np.array([1, 1, 1]),           # 改为 1-30, 1-40, 1-2100
    'MAX_VALUES': np.array([30, 40, 2100]),      # 原 [128, 200, 255]
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 3,
    'PPO_BATCH_SIZE': 64,
    'PPO_EPOCHS': 4,
    'CLIP_EPSILON': 0.2,
    'GAMMA': 0.99,
    'GAE_LAMBDA': 0.95,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'ACTOR_LR': 3e-4,
    'CRITIC_LR': 3e-4,
    'NUM_RUNS': 20,
    'TOP_K_SAMPLES': 20,
    # 目标路径组（保持不变）
    'TARGET_PATHS': [
        # A1
        {2, 3, 4, 6, 7, 9, 10, 11, 12, 13, 14, 16, 17, 19, 22, 27, 30, 33, 36, 38, 44, 46, 47, 50, 53, 55, 61, 63, 65, 67,
         71, 79, 80, 81, 83, 84, 85, 87, 93, 94},
        # A2
        {3, 4, 6, 7, 9, 10, 11, 12, 13, 14, 16, 17, 19, 22, 27, 30, 33, 36, 38, 44, 46, 47, 50, 53, 55, 61, 63, 65, 67, 71,
         79, 80, 81, 83, 84, 85, 87, 93, 94},
        # A3
        {2, 3, 4, 6, 7, 9, 10, 11, 12, 15, 19, 22, 27, 28, 30, 33, 36, 38, 44, 46, 47, 48, 50, 53, 55, 61, 63, 65, 67, 71,
         79, 80, 81, 83, 84, 86, 87, 93, 94},
        # A4
        {3, 4, 5, 7, 9, 10, 11, 12, 13, 16, 17, 19, 22, 24, 27, 30, 33, 36, 38, 44, 45, 47, 50, 53, 55, 61, 63, 65, 67, 71,
         80, 81, 83, 84, 85, 87, 93, 94},
        # A5
        {2, 3, 4, 5, 19, 20, 22, 24, 27, 28, 30, 33, 36, 38, 39, 41, 45, 50, 53, 55, 61, 63, 65, 67, 68, 71, 83, 84, 85, 87,
         90, 91},
        # A6
        {3, 4, 5, 18, 20, 22, 24, 27, 28, 30, 33, 35, 38, 39, 41, 45, 50, 53, 55, 61, 63, 64, 67, 68, 71, 83, 84, 85, 87,
         90, 91},
        # A7
        {3, 4, 8, 19, 22, 23, 25, 30, 33, 36, 38, 40, 42, 43, 49, 53, 55, 61, 63, 65, 67, 69, 71, 79, 80, 81, 82, 92, 93,
         94},
        # A8
        {3, 4, 8, 19, 22, 23, 25, 30, 33, 36, 38, 40, 42, 43, 50, 51, 53, 55, 61, 63, 65, 67, 71, 79, 80, 81, 82, 92, 93,
         94},
        # A9
        {12, 13, 14, 16, 19, 22, 24, 27, 30, 32, 34, 44, 45, 47, 50, 53, 55, 61, 63, 65, 67, 71, 80, 81, 86, 88, 89, 94,
         96},
        # A10
        {1, 12, 15, 19, 22, 24, 27, 28, 30, 33, 36, 38, 44, 45, 47, 48, 50, 53, 55, 61, 63, 65, 67, 71, 80, 81, 86, 93, 94},
        # A11
        {2, 3, 4, 18, 20, 22, 23, 24, 26, 30, 33, 35, 38, 39, 40, 41, 42, 50, 51, 53, 55, 61, 63, 64, 67, 68, 71, 90, 91},
        # A12
        {12, 13, 14, 16, 17, 19, 22, 24, 27, 30, 32, 44, 45, 47, 50, 53, 55, 61, 63, 65, 67, 71, 80, 81, 86, 88, 96, 97},
        # A13
        {12, 13, 14, 16, 17, 19, 22, 24, 27, 30, 32, 44, 45, 47, 50, 53, 55, 58, 65, 67, 71, 80, 81, 86, 88, 96, 97},
        # A14
        {12, 13, 14, 16, 17, 19, 22, 27, 30, 32, 44, 46, 47, 50, 53, 55, 59, 65, 67, 71, 79, 80, 81, 86, 88, 96, 97},
        # A15
        {3, 4, 19, 21, 23, 24, 25, 30, 33, 36, 37, 40, 42, 49, 53, 55, 61, 63, 65, 66, 69, 71, 91},
        # A16
        {2, 3, 4, 8, 21, 30, 33, 37, 49, 53, 55, 56, 61, 63, 66, 71, 78, 79, 80, 81, 82, 95, 97},
        # A17
        {2, 3, 4, 5, 18, 29, 33, 35, 50, 52, 60, 70, 73, 74, 75, 76, 83, 84, 85, 87, 98},
        # A18
        {2, 3, 4, 8, 21, 30, 33, 37, 49, 54, 62, 71, 74, 78, 79, 80, 81, 82, 95, 97},
        # A19
        {2, 3, 4, 18, 30, 31, 33, 35, 50, 51, 53, 55, 57, 61, 63, 64, 71, 99, 100},
        # A20
        {2, 3, 4, 8, 21, 30, 33, 37, 49, 54, 62, 72, 77, 79, 80, 81, 82, 95, 97}
    ]
}

# === 执行规则函数（未修改） ===
def execute_Tr(a):
    cpu_cores, memory_gb, disk_space_gb = int(a[0]), float(a[1]), float(a[2])
    triggered = set()
    b = {}
    if (cpu_cores >= 16) != (cpu_cores >= 13):
        b[0] = 1; triggered.add(1)
    if (cpu_cores >= 16) != (cpu_cores >= 18):
        b[1] = 2; triggered.add(2)
    if (cpu_cores >= 16) != (cpu_cores >= 21):
        b[2] = 3; triggered.add(3)
    if (cpu_cores >= 16) != (cpu_cores >= 24):
        b[3] = 4; triggered.add(4)
    if cpu_cores >= 16:
        if (memory_gb >= 32.0 and disk_space_gb >= 1000) != (memory_gb != 32.0 and disk_space_gb >= 1000):
            b[4] = 5; triggered.add(5)
        if (memory_gb >= 32.0 and disk_space_gb >= 1000) != (memory_gb == 32.0 and disk_space_gb >= 1000):
            b[5] = 6; triggered.add(6)
        if (memory_gb >= 32.0 and disk_space_gb >= 1000) != (memory_gb >= 32.0 and disk_space_gb == 1000):
            b[6] = 7; triggered.add(7)
        if (memory_gb >= 32.0 and disk_space_gb >= 1000) != (memory_gb >= 32.0 and disk_space_gb != 1000):
            b[7] = 8; triggered.add(8)
        if (memory_gb >= 32.0 and disk_space_gb >= 1000) != (memory_gb >= 38.0 and disk_space_gb >= 1000):
            b[8] = 9; triggered.add(9)
        if (memory_gb >= 32.0 and disk_space_gb >= 1000) != (memory_gb >= 32.0 and disk_space_gb >= 1750):
            b[9] = 10; triggered.add(10)
        if (memory_gb >= 32.0 and disk_space_gb >= 1000) != (memory_gb >= 45.0 and disk_space_gb >= 1000):
            b[10] = 11; triggered.add(11)
    if memory_gb >= 32.0 and disk_space_gb >= 1000:
        if (disk_space_gb >= 2000) != (disk_space_gb != 2000):
            b[11] = 12; triggered.add(12)
        if (disk_space_gb >= 2000) != (disk_space_gb >= 1540):
            b[12] = 13; triggered.add(13)
        if (disk_space_gb >= 2000) != (disk_space_gb >= 1670):
            b[13] = 14; triggered.add(14)
        if (disk_space_gb >= 1500) != (disk_space_gb != 1500):
            b[14] = 15; triggered.add(15)
        if (disk_space_gb >= 1500) != (disk_space_gb == 1500):
            b[15] = 16; triggered.add(16)
        if (disk_space_gb >= 1500) != (disk_space_gb >= 1770):
            b[16] = 17; triggered.add(17)
    if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb != 16.0 and disk_space_gb >= 500):
        b[17] = 18; triggered.add(18)
    if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb == 16.0 and disk_space_gb >= 500):
        b[18] = 19; triggered.add(19)
    if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 26.0 and disk_space_gb >= 500):
        b[19] = 20; triggered.add(20)
    if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb != 500):
        b[20] = 21; triggered.add(21)
    if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb == 500):
        b[21] = 22; triggered.add(22)
    if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb >= 900):
        b[22] = 23; triggered.add(23)
    if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 33.0 and disk_space_gb >= 500):
        b[23] = 24; triggered.add(24)
    if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb >= 780):
        b[24] = 25; triggered.add(25)
    elif memory_gb >= 16.0 and disk_space_gb >= 500:
        if (disk_space_gb >= 1000) != (disk_space_gb != 1000):
            b[25] = 26; triggered.add(26)
        if (disk_space_gb >= 1000) != (disk_space_gb == 1000):
            b[26] = 27; triggered.add(27)
        if (disk_space_gb >= 1000) != (disk_space_gb >= 1200):
            b[27] = 28; triggered.add(28)
    if (memory_gb >= 8.0) != (memory_gb != 8.0):
        b[28] = 29; triggered.add(29)
    if (memory_gb >= 8.0) != (memory_gb == 8.0):
        b[29] = 30; triggered.add(30)
    if (memory_gb >= 8.0) != (memory_gb >= 12.0):
        b[30] = 31; triggered.add(31)
    if (cpu_cores >= 8) != (cpu_cores != 8):
        b[31] = 32; triggered.add(32)
    if (cpu_cores >= 8) != (cpu_cores == 8):
        b[32] = 33; triggered.add(33)
    if (cpu_cores >= 8) != (cpu_cores >= 9):
        b[33] = 34; triggered.add(34)
    elif cpu_cores >= 8:
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb != 16.0 and disk_space_gb >= 500):
            b[34] = 35; triggered.add(35)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb == 16.0 and disk_space_gb >= 500):
            b[35] = 36; triggered.add(36)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb != 500):
            b[36] = 37; triggered.add(37)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb == 500):
            b[37] = 38; triggered.add(38)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 22.0 and disk_space_gb >= 500):
            b[38] = 39; triggered.add(39)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb >= 800):
            b[39] = 40; triggered.add(40)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 27.0 and disk_space_gb >= 500):
            b[40] = 41; triggered.add(41)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb >= 940):
            b[41] = 42; triggered.add(42)
    if memory_gb >= 16.0 and disk_space_gb >= 500:
        if (disk_space_gb >= 1000 and memory_gb >= 32.0) != (disk_space_gb != 1000 and memory_gb >= 32.0):
            b[42] = 43; triggered.add(43)
        if (disk_space_gb >= 1000 and memory_gb >= 32.0) != (disk_space_gb == 1000 and memory_gb >= 32.0):
            b[43] = 44; triggered.add(44)
        if (disk_space_gb >= 1000 and memory_gb >= 32.0) != (disk_space_gb >= 1000 and memory_gb != 32.0):
            b[44] = 45; triggered.add(45)
        if (disk_space_gb >= 1000 and memory_gb >= 32.0) != (disk_space_gb >= 1000 and memory_gb == 32.0):
            b[45] = 46; triggered.add(46)
        if (disk_space_gb >= 1000 and memory_gb >= 32.0) != (disk_space_gb >= 1000 and memory_gb >= 36.0):
            b[46] = 47; triggered.add(47)
        if (disk_space_gb >= 1000 and memory_gb >= 32.0) != (disk_space_gb >= 1100 and memory_gb >= 32.0):
            b[47] = 48; triggered.add(48)
    if (disk_space_gb >= 750) != (disk_space_gb != 750):
        b[48] = 49; triggered.add(49)
    if (disk_space_gb >= 750) != (disk_space_gb == 750):
        b[49] = 50; triggered.add(50)
    if (disk_space_gb >= 750) != (disk_space_gb >= 850):
        b[50] = 51; triggered.add(51)
    if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb != 8.0 and disk_space_gb >= 250):
        b[51] = 52; triggered.add(52)
    if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb == 8.0 and disk_space_gb >= 250):
        b[52] = 53; triggered.add(53)
    if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb >= 8.0 and disk_space_gb != 250):
        b[53] = 54; triggered.add(54)
    if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb >= 8.0 and disk_space_gb == 250):
        b[54] = 55; triggered.add(55)
    if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb >= 8.0 and disk_space_gb >= 280):
        b[55] = 56; triggered.add(56)
    if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb >= 10.0 and disk_space_gb >= 250):
        b[56] = 57; triggered.add(57)
    if (cpu_cores >= 4) != (cpu_cores != 4):
        b[57] = 58; triggered.add(58)
    if (cpu_cores >= 4) != (cpu_cores >= 6):
        b[58] = 59; triggered.add(59)
    elif cpu_cores >= 4:
        if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb != 9.5 and disk_space_gb >= 250):
            b[59] = 60; triggered.add(60)
        if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb == 8.0 and disk_space_gb >= 250):
            b[60] = 61; triggered.add(61)
        if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb >= 8.0 and disk_space_gb != 250):
            b[61] = 62; triggered.add(62)
        if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb >= 8.0 and disk_space_gb == 250):
            b[62] = 63; triggered.add(63)
    if memory_gb >= 8.0 and disk_space_gb >= 250:
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb != 16.0 and disk_space_gb >= 500):
            b[63] = 64; triggered.add(64)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb == 16.0 and disk_space_gb >= 500):
            b[64] = 65; triggered.add(65)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb != 500):
            b[65] = 66; triggered.add(66)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb == 500):
            b[66] = 67; triggered.add(67)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 19.0 and disk_space_gb >= 500):
            b[67] = 68; triggered.add(68)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb >= 660):
            b[68] = 69; triggered.add(69)
    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb != 4.0 and disk_space_gb >= 100):
        b[69] = 70; triggered.add(70)
    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb == 4.0 and disk_space_gb >= 100):
        b[70] = 71; triggered.add(71)
    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb >= 4.0 and disk_space_gb != 100):
        b[71] = 72; triggered.add(72)
    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb >= 6.2 and disk_space_gb >= 100):
        b[72] = 73; triggered.add(73)
    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb >= 7.0 and disk_space_gb >= 210):
        b[73] = 74; triggered.add(74)
    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb != 4.0 and disk_space_gb >= 100):
        b[74] = 75; triggered.add(75)
    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb >= 5.3 and disk_space_gb >= 100):
        b[75] = 76; triggered.add(76)
    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb >= 4.0 and disk_space_gb != 100):
        b[76] = 77; triggered.add(77)
    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb >= 4.0 and disk_space_gb >= 278):
        b[77] = 78; triggered.add(78)
    if (memory_gb >= 32.0) != (memory_gb == 32.0):
        b[78] = 79; triggered.add(79)
    if (memory_gb >= 32.0) != (memory_gb >= 36.5):
        b[79] = 80; triggered.add(80)
    if (memory_gb >= 32.0) != (memory_gb >= 35.0):
        b[80] = 81; triggered.add(81)
    if memory_gb >= 32.0:
        if (cpu_cores >= 16 and disk_space_gb >= 1000) != (cpu_cores >= 16 and disk_space_gb != 1000):
            b[81] = 82; triggered.add(82)
    if (cpu_cores >= 16 and disk_space_gb >= 1000) != (cpu_cores >= 16 and disk_space_gb == 1000):
        b[82] = 83; triggered.add(83)
    if (cpu_cores >= 16 and disk_space_gb >= 1000) != (cpu_cores >= 22 and disk_space_gb >= 1000):
        b[83] = 84; triggered.add(84)
    if (cpu_cores >= 16 and disk_space_gb >= 1000) != (cpu_cores == 16 and disk_space_gb >= 1000):
        b[84] = 85; triggered.add(85)
    if (cpu_cores >= 16 and disk_space_gb >= 1000) != (cpu_cores != 16 and disk_space_gb >= 1000):
        b[85] = 86; triggered.add(86)
    if (cpu_cores >= 16 and disk_space_gb >= 1000) != (cpu_cores >= 31 and disk_space_gb >= 1000):
        b[86] = 87; triggered.add(87)
    if (cpu_cores >= 8) != (cpu_cores != 8):
        b[87] = 88; triggered.add(88)
    if (cpu_cores >= 8) != (cpu_cores >= 9.4):
        b[88] = 89; triggered.add(89)
    if (memory_gb >= 16.0) != (memory_gb >= 22):
        b[89] = 90; triggered.add(90)
    if (memory_gb >= 16.0) != (memory_gb >= 31):
        b[90] = 91; triggered.add(91)
    elif memory_gb >= 16.0:
        if (cpu_cores >= 8 and disk_space_gb >= 500) != (cpu_cores >= 8 and disk_space_gb >= 950):
            b[91] = 92; triggered.add(92)
        if (cpu_cores >= 8 and disk_space_gb >= 500) != (cpu_cores == 8 and disk_space_gb >= 500):
            b[92] = 93; triggered.add(93)
        if (cpu_cores >= 8 and disk_space_gb >= 500) != (cpu_cores >= 8 and disk_space_gb == 500):
            b[93] = 94; triggered.add(94)
        if (cpu_cores >= 8 and disk_space_gb >= 500) != (cpu_cores >= 8 and disk_space_gb != 500):
            b[94] = 95; triggered.add(95)
        if (cpu_cores >= 8 and disk_space_gb >= 500) != (cpu_cores != 8 and disk_space_gb >= 500):
            b[95] = 96; triggered.add(96)
        if (cpu_cores >= 8 and disk_space_gb >= 500) != (cpu_cores >= 8 or disk_space_gb >= 500):
            b[96] = 97; triggered.add(97)
    if (memory_gb >= 8.0) != (memory_gb != 8.0):
        b[97] = 98; triggered.add(98)
    if (memory_gb >= 8.0) != (memory_gb >= 11):
        b[98] = 99; triggered.add(99)
    if (memory_gb >= 8.0) != (memory_gb >= 12.0):
        b[99] = 100; triggered.add(100)
    return triggered

# === 状态处理辅助函数 ===
def clip_state(state):
    return np.clip(state, EXPERIMENT_CONFIG['MIN_VALUES'], EXPERIMENT_CONFIG['MAX_VALUES'])

def denormalize_state(normalized_state):
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return normalized_state * (max_vals - min_vals) / 2 + (min_vals + max_vals) / 2

def coverage_similarity(triggered, target_path):
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0
    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)

def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']
    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']
    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']
    return reward

# === PPO Actor ===
class PPOActor(nn.Module):
    def __init__(self, state_dim=3, action_dim=3, hidden_dim=256):
        super(PPOActor, self).__init__()
        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)
        self.mean_head = nn.Linear(hidden_dim, action_dim)
        self.log_std_head = nn.Linear(hidden_dim, action_dim)
        self.action_scale = 10.0

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))
        x = F.relu(self.fc3(x))
        mean = torch.tanh(self.mean_head(x)) * self.action_scale
        log_std = torch.clamp(self.log_std_head(x), -20, 2)
        return mean, log_std

    def get_action_and_log_prob(self, state):
        mean, log_std = self.forward(state)
        std = torch.exp(log_std)
        dist = torch.distributions.Normal(mean, std)
        action = dist.sample()
        log_prob = dist.log_prob(action).sum(dim=-1)
        return action, log_prob

    def get_log_prob(self, state, action):
        mean, log_std = self.forward(state)
        std = torch.exp(log_std)
        dist = torch.distributions.Normal(mean, std)
        log_prob = dist.log_prob(action).sum(dim=-1)
        return log_prob

# === PPO Critic ===
class PPOCritic(nn.Module):
    def __init__(self, state_dim=3, hidden_dim=256):
        super(PPOCritic, self).__init__()
        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)
        self.value_head = nn.Linear(hidden_dim, 1)

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))
        x = F.relu(self.fc3(x))
        value = self.value_head(x).squeeze(-1)
        return value

# === PPO Buffer ===
class PPOBuffer:
    def __init__(self):
        self.states = []
        self.actions = []
        self.rewards = []
        self.values = []
        self.log_probs = []
        self.dones = []
        self.path_indices = []
        self.similarities = []

    def store(self, state, action, reward, value, log_prob, done, path_idx, similarity):
        self.states.append(state)
        self.actions.append(action)
        self.rewards.append(reward)
        self.values.append(value)
        self.log_probs.append(log_prob)
        self.dones.append(done)
        self.path_indices.append(path_idx)
        self.similarities.append(similarity)

    def compute_advantages(self):
        rewards = np.array(self.rewards)
        values = np.array(self.values)
        dones = np.array(self.dones)
        advantages = np.zeros_like(rewards)
        last_advantage = 0
        for t in reversed(range(len(rewards))):
            if t == len(rewards) - 1:
                next_value = 0
            else:
                next_value = values[t + 1]
            delta = rewards[t] + EXPERIMENT_CONFIG['GAMMA'] * next_value * (1 - dones[t]) - values[t]
            advantages[t] = delta + EXPERIMENT_CONFIG['GAMMA'] * EXPERIMENT_CONFIG['GAE_LAMBDA'] * (
                    1 - dones[t]) * last_advantage
            last_advantage = advantages[t]
        returns = advantages + values
        advantages = (advantages - np.mean(advantages)) / (np.std(advantages) + 1e-8)
        return advantages, returns

    def get_batch_iterator(self, advantages, returns):
        indices = np.arange(len(self.states))
        np.random.shuffle(indices)
        batch_size = EXPERIMENT_CONFIG['PPO_BATCH_SIZE']
        for start in range(0, len(self.states), batch_size):
            end = min(start + batch_size, len(self.states))
            batch_indices = indices[start:end]
            yield {
                'states': torch.FloatTensor(np.array([self.states[i] for i in batch_indices])).to(device),
                'actions': torch.FloatTensor(np.array([self.actions[i] for i in batch_indices])).to(device),
                'old_log_probs': torch.FloatTensor(np.array([self.log_probs[i] for i in batch_indices])).to(device),
                'advantages': torch.FloatTensor(advantages[batch_indices]).to(device),
                'returns': torch.FloatTensor(returns[batch_indices]).to(device)
            }

    def get_top_k_per_path(self, num_paths, k=20):
        path_samples = {i: [] for i in range(num_paths)}
        for idx in range(len(self.states)):
            path_idx = self.path_indices[idx]
            similarity = self.similarities[idx]
            state = self.states[idx]
            path_samples[path_idx].append((idx, similarity, state))
        top_k_results = {}
        for path_idx in range(num_paths):
            samples = path_samples[path_idx]
            if len(samples) == 0:
                top_k_results[path_idx] = []
                continue
            samples.sort(key=lambda x: x[1], reverse=True)
            top_k = samples[:k]
            top_k_results[path_idx] = []
            for sample in top_k:
                normalized_state = sample[2]
                original_state = denormalize_state(normalized_state)
                original_state_int = np.round(original_state).astype(int)
                top_k_results[path_idx].append({
                    'state': original_state_int,
                    'similarity': sample[1],
                    'triggered': execute_Tr(*original_state_int)
                })
        return top_k_results

    def clear(self):
        self.states.clear()
        self.actions.clear()
        self.rewards.clear()
        self.values.clear()
        self.log_probs.clear()
        self.dones.clear()
        self.path_indices.clear()
        self.similarities.clear()

    def __len__(self):
        return len(self.states)

# === PPO Agent ===
class PPOAgent:
    def __init__(self, state_dim=3, action_dim=3):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.actor = PPOActor(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.actor_optimizer = optim.Adam(self.actor.parameters(), lr=EXPERIMENT_CONFIG['ACTOR_LR'])
        self.critic = PPOCritic(state_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_optimizer = optim.Adam(self.critic.parameters(), lr=EXPERIMENT_CONFIG['CRITIC_LR'])
        self.buffer = PPOBuffer()
        self.update_count = 0

    def get_action(self, state):
        min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
        max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
        normalized_state = (state - (min_vals + max_vals) / 2) / ((max_vals - min_vals) / 2)
        state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)
        with torch.no_grad():
            action, log_prob = self.actor.get_action_and_log_prob(state_tensor)
            value = self.critic(state_tensor)
        action = action.cpu().numpy()[0]
        log_prob = log_prob.cpu().item()
        value = value.cpu().item()
        return action, log_prob, value

    def store_experience(self, state, action, reward, value, log_prob, done, path_idx, similarity):
        min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
        max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
        normalized_state = (state - (min_vals + max_vals) / 2) / ((max_vals - min_vals) / 2)
        self.buffer.store(normalized_state, action, reward, value, log_prob, done, path_idx, similarity)

    def update(self):
        if len(self.buffer) == 0:
            return
        advantages, returns = self.buffer.compute_advantages()
        for epoch in range(EXPERIMENT_CONFIG['PPO_EPOCHS']):
            for batch in self.buffer.get_batch_iterator(advantages, returns):
                new_log_probs = self.actor.get_log_prob(batch['states'], batch['actions'])
                ratio = torch.exp(new_log_probs - batch['old_log_probs'])
                surr1 = ratio * batch['advantages']
                surr2 = torch.clamp(ratio, 1 - EXPERIMENT_CONFIG['CLIP_EPSILON'],
                                    1 + EXPERIMENT_CONFIG['CLIP_EPSILON']) * batch['advantages']
                actor_loss = -torch.min(surr1, surr2).mean()
                self.actor_optimizer.zero_grad()
                actor_loss.backward()
                torch.nn.utils.clip_grad_norm_(self.actor.parameters(), 0.5)
                self.actor_optimizer.step()
                new_values = self.critic(batch['states'])
                critic_loss = F.mse_loss(new_values, batch['returns'])
                self.critic_optimizer.zero_grad()
                critic_loss.backward()
                torch.nn.utils.clip_grad_norm_(self.critic.parameters(), 0.5)
                self.critic_optimizer.step()
        self.update_count += 1
        self.buffer.clear()

# === 性能指标计算 ===
def calculate_run_performance(run_idx, ppo_results, training_time, total_steps, update_count, agent):
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)
    total_reward = 0
    average_reward = 0
    convergence = 0
    environment_adaptability = 0
    generalization_ability = 0
    computational_efficiency = 0
    policy_update_frequency = 0
    all_similarities = []
    total_samples = 0
    all_rewards = []
    for path_idx in range(num_paths):
        samples = ppo_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']
            total_reward += reward
            all_rewards.append(reward)
            all_similarities.append(similarity)
            total_samples += 1
    if total_samples > 0:
        average_reward = total_reward / total_samples
    if all_similarities:
        convergence = np.mean(all_similarities)
    if len(all_similarities) > 1:
        environment_adaptability = 1 / (np.std(all_similarities) + 1e-8)
    generalization_ability = convergence
    if training_time > 0:
        computational_efficiency = total_steps / training_time
    if training_time > 0:
        policy_update_frequency = update_count / training_time
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0
    return {
        'Run': run_idx + 1,
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }

# === Excel 导出 ===
def export_to_excel(all_ppo_results, all_performance_data, target_paths, output_path="PPO_20_run.xlsx"):
    print("\n生成Excel...")
    all_ppo_summary_data = []
    all_ppo_detailed_data = []
    for run_idx, (ppo_results, performance_data) in enumerate(zip(all_ppo_results, all_performance_data)):
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = ppo_results[path_idx]
            if len(samples) == 0:
                all_ppo_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue
            similarities = [s['similarity'] for s in samples]
            is_perfect = 'Yes' if any(abs(s - 1.0) < 0.001 for s in similarities) else 'No'
            all_ppo_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })
            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']
                all_ppo_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Sample ID': sample_idx + 1,
                    'X': int(state[0]),
                    'Y': int(state[1]),
                    'Z': int(state[2]),
                    'Similarity': round(similarity, 4),
                    'Perfect Coverage': 'Yes' if abs(similarity - 1.0) < 0.001 else 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path))),
                    'Triggered Rules': ', '.join(map(str, sorted(triggered))),
                    'Hit Rule Count': len(target_path.intersection(triggered)),
                    'Target Rule Count': len(target_path)
                })

    ppo_summary_df = pd.DataFrame(all_ppo_summary_data)
    ppo_detailed_df = pd.DataFrame(all_ppo_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        ppo_summary_df.to_excel(writer, sheet_name='PPO路径摘要', index=False)
        ppo_detailed_df.to_excel(writer, sheet_name='PPO详细样本数据', index=False)
        selected_columns = ['Run', 'Total Reward', 'Average Reward', 'Convergence',
                            'Environment Adaptability', 'Generalization Ability',
                            'Computational Efficiency', 'Policy Update Frequency',
                            'Average Similarity', 'Max Similarity', 'Min Similarity']
        performance_df[selected_columns].to_excel(writer, sheet_name='性能指标', index=False)

        workbook = writer.book
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        perfect_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        ws1 = writer.sheets['PPO路径摘要']
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row_idx in range(2, ws1.max_row + 1):
            if ws1.cell(row_idx, 9).value == 'Yes':
                for col_idx in range(1, ws1.max_column + 1):
                    ws1.cell(row_idx, col_idx).fill = perfect_fill
        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 15
        ws1.column_dimensions['F'].width = 15
        ws1.column_dimensions['G'].width = 15
        ws1.column_dimensions['H'].width = 15
        ws1.column_dimensions['I'].width = 15
        ws1.column_dimensions['J'].width = 50

        ws2 = writer.sheets['PPO详细样本数据']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 15
        ws2.column_dimensions['I'].width = 40
        ws2.column_dimensions['J'].width = 40
        ws2.column_dimensions['K'].width = 15
        ws2.column_dimensions['L'].width = 15

        ws3 = writer.sheets['性能指标']
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws3.column_dimensions[col].width = 18

    print(f"Excel已保存: {output_path}")
    print(f"  - Sheet1: PPO路径摘要 ({len(all_ppo_summary_data)} 行)")
    print(f"  - Sheet2: PPO详细样本数据 ({len(all_ppo_detailed_data)} 行)")
    print(f"  - Sheet3: 性能指标 ({len(all_performance_data)} 行)")

# === 训练流程 ===
def train_ppo_workflow():
    print("=" * 80)
    print("PPO训练开始")
    print("奖励基于覆盖相似度")
    print("=" * 80)

    agent = PPOAgent()
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n生成初始样本: 每路径 {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']} 个")
    path_samples = {}
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']

    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            state = np.array([
                np.random.randint(min_vals[0], max_vals[0] + 1),
                np.random.randint(min_vals[1], max_vals[1] + 1),
                np.random.randint(min_vals[2], max_vals[2] + 1)
            ], dtype=np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  路径 {path_idx + 1}/{num_paths}: 已生成 {len(samples)} 个样本")

    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size

    print(f"\n训练配置: 每批 {batch_size} 个样本, 每样本 {EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']} 步")
    print(f"总批次: {num_batches * num_paths} 批")
    print("-" * 80)

    global_buffer = PPOBuffer()

    for batch_idx in range(num_batches):
        print(f"\n批次 {batch_idx + 1}/{num_batches}")
        for path_idx in range(num_paths):
            target_path = target_paths[path_idx]
            batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]
            batch_rewards = []
            batch_similarities = []

            for sample_idx, initial_state in enumerate(batch_samples):
                state = initial_state.copy()
                episode_reward = 0
                final_similarity = 0

                for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                    action, log_prob, value = agent.get_action(state)
                    next_state = state + action
                    next_state = clip_state(next_state)
                    triggered = execute_Tr(*next_state)
                    reward = unified_reward_function(triggered, target_path)
                    similarity = coverage_similarity(triggered, target_path)
                    done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)

                    agent.store_experience(state, action, reward, value, log_prob, done, path_idx, similarity)
                    normalized_state = (state - (min_vals + max_vals) / 2) / ((max_vals - min_vals) / 2)
                    global_buffer.store(normalized_state, action, reward, value, log_prob, done, path_idx, similarity)

                    state = next_state
                    episode_reward += reward
                    final_similarity = similarity
                    total_steps += 1

                batch_rewards.append(episode_reward)
                batch_similarities.append(final_similarity)

            avg_reward = np.mean(batch_rewards)
            avg_similarity = np.mean(batch_similarities)
            print(f"  路径 {path_idx + 1}: 平均奖励={avg_reward:.2f}, 平均相似度={avg_similarity:.4f}")

        print(f"\n  执行 PPO 更新...")
        agent.update()
        print(f"  全局经验池大小: {len(global_buffer)}")

    training_time = time.time() - start_time
    print("\n" + "=" * 80)
    print(f"PPO训练完成! 总耗时: {training_time:.2f} 秒, 总步数: {total_steps}")
    print(f"全局经验池大小: {len(global_buffer)}")
    print(f"PPO更新次数: {agent.update_count}")
    print("=" * 80)

    print(f"\n提取每路径 Top-{EXPERIMENT_CONFIG['TOP_K_SAMPLES']} 高相似度样本...")
    ppo_top_k_results = global_buffer.get_top_k_per_path(num_paths, EXPERIMENT_CONFIG['TOP_K_SAMPLES'])
    return agent, ppo_top_k_results, training_time, total_steps, agent.update_count

# === 主函数 ===
def main():
    print("\n" + "=" * 80)
    print("PPO - 20次运行实验")
    print("性能指标：覆盖相似度")
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    print(f"状态范围: X[{min_vals[0]}-{max_vals[0]}], Y[{min_vals[1]}-{max_vals[1]}], Z[{min_vals[2]}-{max_vals[2]}]")
    print("=" * 80)

    all_ppo_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'='*80}")
        print(f"开始第 {run_idx + 1}/{EXPERIMENT_CONFIG['NUM_RUNS']} 次运行")
        print(f"{'='*80}")

        ppo_agent, ppo_results, training_time, total_steps, update_count = train_ppo_workflow()
        performance_data = calculate_run_performance(
            run_idx, ppo_results, training_time, total_steps, update_count, ppo_agent
        )

        all_ppo_results.append(ppo_results)
        all_performance_data.append(performance_data)

        print(f"\n第 {run_idx + 1} 次运行完成!")
        print(f"  总奖励: {performance_data['Total Reward']}")
        print(f"  平均奖励: {performance_data['Average Reward']}")
        print(f"  平均相似度: {performance_data['Average Similarity']}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"PPO_20_runs_{timestamp}.xlsx"
    export_to_excel(all_ppo_results, all_performance_data, target_paths, output_path)

    print("\n" + "=" * 80)
    print("20次运行统计汇总")
    print("=" * 80)
    total_rewards = [p['Total Reward'] for p in all_performance_data]
    average_rewards = [p['Average Reward'] for p in all_performance_data]
    convergences = [p['Convergence'] for p in all_performance_data]
    environment_adaptabilities = [p['Environment Adaptability'] for p in all_performance_data]
    generalization_abilities = [p['Generalization Ability'] for p in all_performance_data]
    computational_efficiencies = [p['Computational Efficiency'] for p in all_performance_data]
    policy_update_frequencies = [p['Policy Update Frequency'] for p in all_performance_data]
    avg_similarities = [p['Average Similarity'] for p in all_performance_data]

    print(f"\n总奖励:")
    print(f"  平均值: {np.mean(total_rewards):.2f}")
    print(f"  标准差: {np.std(total_rewards):.2f}")

    print(f"\n平均奖励:")
    print(f"  平均值: {np.mean(average_rewards):.4f}")
    print(f"  标准差: {np.std(average_rewards):.4f}")

    print(f"\n收敛度 (平均相似度):")
    print(f"  平均值: {np.mean(convergences):.4f}")
    print(f"  标准差: {np.std(convergences):.4f}")

    print(f"\n环境适应性:")
    print(f"  平均值: {np.mean(environment_adaptabilities):.4f}")
    print(f"  标准差: {np.std(environment_adaptabilities):.4f}")

    print(f"\n泛化能力:")
    print(f"  平均值: {np.mean(generalization_abilities):.4f}")
    print(f"  标准差: {np.std(generalization_abilities):.4f}")

    print(f"\n计算效率 (步/秒):")
    print(f"  平均值: {np.mean(computational_efficiencies):.2f}")
    print(f"  标准差: {np.std(computational_efficiencies):.2f}")

    print(f"\n策略更新频率 (次/秒):")
    print(f"  平均值: {np.mean(policy_update_frequencies):.4f}")
    print(f"  标准差: {np.std(policy_update_frequencies):.4f}")

    print(f"\n平均相似度统计:")
    print(f"  平均值: {np.mean(avg_similarities):.4f}")
    print(f"  标准差: {np.std(avg_similarities):.4f}")

    print("\n" + "=" * 80)
    print(f"所有 {EXPERIMENT_CONFIG['NUM_RUNS']} 次运行完成!")
    print("=" * 80)

if __name__ == "__main__":
    main()