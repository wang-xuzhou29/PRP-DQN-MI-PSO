import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
from torch.distributions import Normal
import numpy as np
import random
import time
from collections import deque
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# ==================== 资源配置规则专用配置 ====================
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'ACTION_DIM': 3,
    # 测试范围：cpu_cores[1,30], memory_gb[1,40], disk_space_gb[1,2100]
    'MIN_VALUES': np.array([1, 1.0, 1.0], dtype=np.float32),
    'MAX_VALUES': np.array([30, 40.0, 2100.0], dtype=np.float32),
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 3,
    'REPLAY_BATCH_SIZE': 64,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'ACTOR_LR': 3e-4,
    'CRITIC_LR': 3e-4,
    'ALPHA_LR': 3e-4,
    'GAMMA': 0.99,
    'TAU': 0.005,
    'NUM_RUNS': 20,
    'TOP_K_SAMPLES': 20,
    # === 目标路径组 ===
    'TARGET_PATHS': [
        # A1
        {2, 3, 4, 6, 7, 9, 10, 11, 12, 13, 14, 16, 17, 19, 22, 27, 30, 33, 36, 38, 44, 46, 47, 50, 53, 55, 61, 63, 65, 67,
         71, 79, 80, 81, 83, 84, 85, 87, 93, 94},
        # A2
        {3, 4, 6, 7, 9, 10, 11, 12, 13, 14, 16, 17, 19, 22, 27, 30, 33, 36, 38, 44, 46, 47, 50, 53, 55, 61, 63, 65, 67, 71,
         79, 80, 81, 83, 84, 85, 87, 93, 94},
        # A3
        {2, 3, 4, 6, 7, 9, 10, 11, 12, 15, 19, 22, 27, 28, 30, 33, 36, 38, 44, 46, 47, 48, 50, 53, 55, 61, 63, 65, 67, 71,
         79, 80, 81, 83, 84, 86, 87, 93, 94},
        # A4
        {3, 4, 5, 7, 9, 10, 11, 12, 13, 16, 17, 19, 22, 24, 27, 30, 33, 36, 38, 44, 45, 47, 50, 53, 55, 61, 63, 65, 67, 71,
         80, 81, 83, 84, 85, 87, 93, 94},
        # A5
        {2, 3, 4, 5, 19, 20, 22, 24, 27, 28, 30, 33, 36, 38, 39, 41, 45, 50, 53, 55, 61, 63, 65, 67, 68, 71, 83, 84, 85, 87,
         90, 91},
        # A6
        {3, 4, 5, 18, 20, 22, 24, 27, 28, 30, 33, 35, 38, 39, 41, 45, 50, 53, 55, 61, 63, 64, 67, 68, 71, 83, 84, 85, 87,
         90, 91},
        # A7
        {3, 4, 8, 19, 22, 23, 25, 30, 33, 36, 38, 40, 42, 43, 49, 53, 55, 61, 63, 65, 67, 69, 71, 79, 80, 81, 82, 92, 93,
         94},
        # A8
        {3, 4, 8, 19, 22, 23, 25, 30, 33, 36, 38, 40, 42, 43, 50, 51, 53, 55, 61, 63, 65, 67, 71, 79, 80, 81, 82, 92, 93,
         94},
        # A9
        {12, 13, 14, 16, 19, 22, 24, 27, 30, 32, 34, 44, 45, 47, 50, 53, 55, 61, 63, 65, 67, 71, 80, 81, 86, 88, 89, 94,
         96},
        # A10
        {1, 12, 15, 19, 22, 24, 27, 28, 30, 33, 36, 38, 44, 45, 47, 48, 50, 53, 55, 61, 63, 65, 67, 71, 80, 81, 86, 93, 94},
        # A11
        {2, 3, 4, 18, 20, 22, 23, 24, 26, 30, 33, 35, 38, 39, 40, 41, 42, 50, 51, 53, 55, 61, 63, 64, 67, 68, 71, 90, 91},
        # A12
        {12, 13, 14, 16, 17, 19, 22, 24, 27, 30, 32, 44, 45, 47, 50, 53, 55, 61, 63, 65, 67, 71, 80, 81, 86, 88, 96, 97},
        # A13
        {12, 13, 14, 16, 17, 19, 22, 24, 27, 30, 32, 44, 45, 47, 50, 53, 55, 58, 65, 67, 71, 80, 81, 86, 88, 96, 97},
        # A14
        {12, 13, 14, 16, 17, 19, 22, 27, 30, 32, 44, 46, 47, 50, 53, 55, 59, 65, 67, 71, 79, 80, 81, 86, 88, 96, 97},
        # A15
        {3, 4, 19, 21, 23, 24, 25, 30, 33, 36, 37, 40, 42, 49, 53, 55, 61, 63, 65, 66, 69, 71, 91},
        # A16
        {2, 3, 4, 8, 21, 30, 33, 37, 49, 53, 55, 56, 61, 63, 66, 71, 78, 79, 80, 81, 82, 95, 97},
        # A17
        {2, 3, 4, 5, 18, 29, 33, 35, 50, 52, 60, 70, 73, 74, 75, 76, 83, 84, 85, 87, 98},
        # A18
        {2, 3, 4, 8, 21, 30, 33, 37, 49, 54, 62, 71, 74, 78, 79, 80, 81, 82, 95, 97},
        # A19
        {2, 3, 4, 18, 30, 31, 33, 35, 50, 51, 53, 55, 57, 61, 63, 64, 71, 99, 100},
        # A20
        {2, 3, 4, 8, 21, 30, 33, 37, 49, 54, 62, 72, 77, 79, 80, 81, 82, 95, 97}
    ]
}

# ==================== 状态处理辅助函数 ====================
def clip_state(state):
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return np.clip(state, min_vals, max_vals)


def normalize_state(state):
    """将原始状态归一化到[-1, 1]区间"""
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return 2 * (state - min_vals) / (max_vals - min_vals) - 1


def denormalize_state(normalized_state):
    """将归一化状态还原为原始值"""
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return (normalized_state + 1) * (max_vals - min_vals) / 2 + min_vals


def coverage_similarity(triggered, target_path):
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0
    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)


def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']
    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']
    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']
    return reward


# ==================== 执行资源配置规则函数 ====================
def execute_Tr(a):
    """
    资源配置规则执行函数
    参数a: 包含3个元素的数组，对应 cpu_cores, memory_gb, disk_space_gb
    返回: 触发的规则编号集合
    """
    cpu_cores, memory_gb, disk_space_gb = int(a[0]), float(a[1]), float(a[2])
    triggered = set()
    b = {}

    # CPU核心数 - 16核层级分支
    if (cpu_cores >= 16) != (cpu_cores >= 13):
        b[0] = 1
        triggered.add(1)
    if (cpu_cores >= 16) != (cpu_cores >= 18):
        b[1] = 2
        triggered.add(2)
    if (cpu_cores >= 16) != (cpu_cores >= 21):
        b[2] = 3
        triggered.add(3)
    if (cpu_cores >= 16) != (cpu_cores >= 24):
        b[3] = 4
        triggered.add(4)

    if cpu_cores >= 16:
        if (memory_gb >= 32.0 and disk_space_gb >= 1000) != (memory_gb != 32.0 and disk_space_gb >= 1000):
            b[4] = 5
            triggered.add(5)
        if (memory_gb >= 32.0 and disk_space_gb >= 1000) != (memory_gb == 32.0 and disk_space_gb >= 1000):
            b[5] = 6
            triggered.add(6)
        if (memory_gb >= 32.0 and disk_space_gb >= 1000) != (memory_gb >= 32.0 and disk_space_gb == 1000):
            b[6] = 7
            triggered.add(7)
        if (memory_gb >= 32.0 and disk_space_gb >= 1000) != (memory_gb >= 32.0 and disk_space_gb != 1000):
            b[7] = 8
            triggered.add(8)
        if (memory_gb >= 32.0 and disk_space_gb >= 1000) != (memory_gb >= 38.0 and disk_space_gb >= 1000):
            b[8] = 9
            triggered.add(9)
        if (memory_gb >= 32.0 and disk_space_gb >= 1000) != (memory_gb >= 32.0 and disk_space_gb >= 1750):
            b[9] = 10
            triggered.add(10)
        if (memory_gb >= 32.0 and disk_space_gb >= 1000) != (memory_gb >= 45.0 and disk_space_gb >= 1000):
            b[10] = 11
            triggered.add(11)

    if memory_gb >= 32.0 and disk_space_gb >= 1000:
        if (disk_space_gb >= 2000) != (disk_space_gb != 2000):
            b[11] = 12
            triggered.add(12)
        if (disk_space_gb >= 2000) != (disk_space_gb >= 1540):
            b[12] = 13
            triggered.add(13)
        if (disk_space_gb >= 2000) != (disk_space_gb >= 1670):
            b[13] = 14
            triggered.add(14)
        if (disk_space_gb >= 1500) != (disk_space_gb != 1500):
            b[14] = 15
            triggered.add(15)
        if (disk_space_gb >= 1500) != (disk_space_gb == 1500):
            b[15] = 16
            triggered.add(16)
        if (disk_space_gb >= 1500) != (disk_space_gb >= 1770):
            b[16] = 17
            triggered.add(17)

    if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb != 16.0 and disk_space_gb >= 500):
        b[17] = 18
        triggered.add(18)
    if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb == 16.0 and disk_space_gb >= 500):
        b[18] = 19
        triggered.add(19)
    if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 26.0 and disk_space_gb >= 500):
        b[19] = 20
        triggered.add(20)
    if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb != 500):
        b[20] = 21
        triggered.add(21)
    if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb == 500):
        b[21] = 22
        triggered.add(22)
    if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb >= 900):
        b[22] = 23
        triggered.add(23)
    if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 33.0 and disk_space_gb >= 500):
        b[23] = 24
        triggered.add(24)
    if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb >= 780):
        b[24] = 25
        triggered.add(25)

    elif memory_gb >= 16.0 and disk_space_gb >= 500:
        if (disk_space_gb >= 1000) != (disk_space_gb != 1000):
            b[25] = 26
            triggered.add(26)
        if (disk_space_gb >= 1000) != (disk_space_gb == 1000):
            b[26] = 27
            triggered.add(27)
        if (disk_space_gb >= 1000) != (disk_space_gb >= 1200):
            b[27] = 28
            triggered.add(28)

    if (memory_gb >= 8.0) != (memory_gb != 8.0):
        b[28] = 29
        triggered.add(29)
    if (memory_gb >= 8.0) != (memory_gb == 8.0):
        b[29] = 30
        triggered.add(30)
    if (memory_gb >= 8.0) != (memory_gb >= 12.0):
        b[30] = 31
        triggered.add(31)

    if (cpu_cores >= 8) != (cpu_cores != 8):
        b[31] = 32
        triggered.add(32)
    if (cpu_cores >= 8) != (cpu_cores == 8):
        b[32] = 33
        triggered.add(33)
    if (cpu_cores >= 8) != (cpu_cores >= 9):
        b[33] = 34
        triggered.add(34)

    elif cpu_cores >= 8:
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb != 16.0 and disk_space_gb >= 500):
            b[34] = 35
            triggered.add(35)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb == 16.0 and disk_space_gb >= 500):
            b[35] = 36
            triggered.add(36)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb != 500):
            b[36] = 37
            triggered.add(37)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb == 500):
            b[37] = 38
            triggered.add(38)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 22.0 and disk_space_gb >= 500):
            b[38] = 39
            triggered.add(39)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb >= 800):
            b[39] = 40
            triggered.add(40)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 27.0 and disk_space_gb >= 500):
            b[40] = 41
            triggered.add(41)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb >= 940):
            b[41] = 42
            triggered.add(42)

    if memory_gb >= 16.0 and disk_space_gb >= 500:
        if (disk_space_gb >= 1000 and memory_gb >= 32.0) != (disk_space_gb != 1000 and memory_gb >= 32.0):
            b[42] = 43
            triggered.add(43)
        if (disk_space_gb >= 1000 and memory_gb >= 32.0) != (disk_space_gb == 1000 and memory_gb >= 32.0):
            b[43] = 44
            triggered.add(44)
        if (disk_space_gb >= 1000 and memory_gb >= 32.0) != (disk_space_gb >= 1000 and memory_gb != 32.0):
            b[44] = 45
            triggered.add(45)
        if (disk_space_gb >= 1000 and memory_gb >= 32.0) != (disk_space_gb >= 1000 and memory_gb == 32.0):
            b[45] = 46
            triggered.add(46)
        if (disk_space_gb >= 1000 and memory_gb >= 32.0) != (disk_space_gb >= 1000 and memory_gb >= 36.0):
            b[46] = 47
            triggered.add(47)
        if (disk_space_gb >= 1000 and memory_gb >= 32.0) != (disk_space_gb >= 1100 and memory_gb >= 32.0):
            b[47] = 48
            triggered.add(48)

    if (disk_space_gb >= 750) != (disk_space_gb != 750):
        b[48] = 49
        triggered.add(49)
    if (disk_space_gb >= 750) != (disk_space_gb == 750):
        b[49] = 50
        triggered.add(50)
    if (disk_space_gb >= 750) != (disk_space_gb >= 850):
        b[50] = 51
        triggered.add(51)

    if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb != 8.0 and disk_space_gb >= 250):
        b[51] = 52
        triggered.add(52)
    if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb == 8.0 and disk_space_gb >= 250):
        b[52] = 53
        triggered.add(53)
    if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb >= 8.0 and disk_space_gb != 250):
        b[53] = 54
        triggered.add(54)
    if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb >= 8.0 and disk_space_gb == 250):
        b[54] = 55
        triggered.add(55)
    if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb >= 8.0 and disk_space_gb >= 280):
        b[55] = 56
        triggered.add(56)
    if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb >= 10.0 and disk_space_gb >= 250):
        b[56] = 57
        triggered.add(57)

    if (cpu_cores >= 4) != (cpu_cores != 4):
        b[57] = 58
        triggered.add(58)
    if (cpu_cores >= 4) != (cpu_cores >= 6):
        b[58] = 59
        triggered.add(59)

    elif cpu_cores >= 4:
        if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb != 9.5 and disk_space_gb >= 250):
            b[59] = 60
            triggered.add(60)
        if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb == 8.0 and disk_space_gb >= 250):
            b[60] = 61
            triggered.add(61)
        if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb >= 8.0 and disk_space_gb != 250):
            b[61] = 62
            triggered.add(62)
        if (memory_gb >= 8.0 and disk_space_gb >= 250) != (memory_gb >= 8.0 and disk_space_gb == 250):
            b[62] = 63
            triggered.add(63)

    if memory_gb >= 8.0 and disk_space_gb >= 250:
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb != 16.0 and disk_space_gb >= 500):
            b[63] = 64
            triggered.add(64)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb == 16.0 and disk_space_gb >= 500):
            b[64] = 65
            triggered.add(65)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb != 500):
            b[65] = 66
            triggered.add(66)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb == 500):
            b[66] = 67
            triggered.add(67)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 19.0 and disk_space_gb >= 500):
            b[67] = 68
            triggered.add(68)
        if (memory_gb >= 16.0 and disk_space_gb >= 500) != (memory_gb >= 16.0 and disk_space_gb >= 660):
            b[68] = 69
            triggered.add(69)

    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb != 4.0 and disk_space_gb >= 100):
        b[69] = 70
        triggered.add(70)
    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb == 4.0 and disk_space_gb >= 100):
        b[70] = 71
        triggered.add(71)
    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb >= 4.0 and disk_space_gb != 100):
        b[71] = 72
        triggered.add(72)
    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb >= 6.2 and disk_space_gb >= 100):
        b[72] = 73
        triggered.add(73)
    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb >= 7.0 and disk_space_gb >= 210):
        b[73] = 74
        triggered.add(74)

    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb != 4.0 and disk_space_gb >= 100):
        b[74] = 75
        triggered.add(75)
    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb >= 5.3 and disk_space_gb >= 100):
        b[75] = 76
        triggered.add(76)
    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb >= 4.0 and disk_space_gb != 100):
        b[76] = 77
        triggered.add(77)
    if (memory_gb >= 4.0 and disk_space_gb >= 100) != (memory_gb >= 4.0 and disk_space_gb >= 278):
        b[77] = 78
        triggered.add(78)

    # 内存分配三维资源分支
    if (memory_gb >= 32.0) != (memory_gb == 32.0):
        b[78] = 79
        triggered.add(79)
    if (memory_gb >= 32.0) != (memory_gb >= 36.5):
        b[79] = 80
        triggered.add(80)
    if (memory_gb >= 32.0) != (memory_gb >= 35.0):
        b[80] = 81
        triggered.add(81)

    if memory_gb >= 32.0:
        if (cpu_cores >= 16 and disk_space_gb >= 1000) != (cpu_cores >= 16 and disk_space_gb != 1000):
            b[81] = 82
            triggered.add(82)

    if (cpu_cores >= 16 and disk_space_gb >= 1000) != (cpu_cores >= 16 and disk_space_gb == 1000):
        b[82] = 83
        triggered.add(83)
    if (cpu_cores >= 16 and disk_space_gb >= 1000) != (cpu_cores >= 22 and disk_space_gb >= 1000):
        b[83] = 84
        triggered.add(84)
    if (cpu_cores >= 16 and disk_space_gb >= 1000) != (cpu_cores == 16 and disk_space_gb >= 1000):
        b[84] = 85
        triggered.add(85)
    if (cpu_cores >= 16 and disk_space_gb >= 1000) != (cpu_cores != 16 and disk_space_gb >= 1000):
        b[85] = 86
        triggered.add(86)
    if (cpu_cores >= 16 and disk_space_gb >= 1000) != (cpu_cores >= 31 and disk_space_gb >= 1000):
        b[86] = 87
        triggered.add(87)

    if (cpu_cores >= 8) != (cpu_cores != 8):
        b[87] = 88
        triggered.add(88)
    if (cpu_cores >= 8) != (cpu_cores >= 9.4):
        b[88] = 89
        triggered.add(89)

    if (memory_gb >= 16.0) != (memory_gb >= 22):
        b[89] = 90
        triggered.add(90)
    if (memory_gb >= 16.0) != (memory_gb >= 31):
        b[90] = 91
        triggered.add(91)

    elif memory_gb >= 16.0:
        if (cpu_cores >= 8 and disk_space_gb >= 500) != (cpu_cores >= 8 and disk_space_gb >= 950):
            b[91] = 92
            triggered.add(92)
        if (cpu_cores >= 8 and disk_space_gb >= 500) != (cpu_cores == 8 and disk_space_gb >= 500):
            b[92] = 93
            triggered.add(93)
        if (cpu_cores >= 8 and disk_space_gb >= 500) != (cpu_cores >= 8 and disk_space_gb == 500):
            b[93] = 94
            triggered.add(94)
        if (cpu_cores >= 8 and disk_space_gb >= 500) != (cpu_cores >= 8 and disk_space_gb != 500):
            b[94] = 95
            triggered.add(95)
        if (cpu_cores >= 8 and disk_space_gb >= 500) != (cpu_cores != 8 and disk_space_gb >= 500):
            b[95] = 96
            triggered.add(96)
        if (cpu_cores >= 8 and disk_space_gb >= 500) != (cpu_cores >= 8 or disk_space_gb >= 500):
            b[96] = 97
            triggered.add(97)

    if (memory_gb >= 8.0) != (memory_gb != 8.0):
        b[97] = 98
        triggered.add(98)
    if (memory_gb >= 8.0) != (memory_gb >= 11):
        b[98] = 99
        triggered.add(99)
    if (memory_gb >= 8.0) != (memory_gb >= 12.0):
        b[99] = 100
        triggered.add(100)

    return triggered


# ==================== SAC Actor ====================
class GaussianPolicy(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dim=256):
        super(GaussianPolicy, self).__init__()
        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.mean = nn.Linear(hidden_dim, action_dim)
        self.log_std = nn.Linear(hidden_dim, action_dim)
        # 适配大范围数值调整动作尺度
        self.action_scale = 100.0
        self.action_bias = 0.0

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))
        mean = self.mean(x)
        log_std = self.log_std(x)
        log_std = torch.clamp(log_std, min=-20, max=2)
        return mean, log_std

    def sample(self, state):
        mean, log_std = self.forward(state)
        std = log_std.exp()
        normal = Normal(mean, std)
        x_t = normal.rsample()
        action = torch.tanh(x_t) * self.action_scale + self.action_bias

        log_prob = normal.log_prob(x_t)
        log_prob -= torch.log(self.action_scale * (1 - torch.tanh(x_t).pow(2)) + 1e-6)
        log_prob = log_prob.sum(1, keepdim=True)

        mean = torch.tanh(mean) * self.action_scale + self.action_bias
        return action, log_prob, mean


# ==================== SAC Critic ====================
class QNetwork(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dim=256):
        super(QNetwork, self).__init__()
        self.fc1 = nn.Linear(state_dim + action_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, 1)

        self.fc4 = nn.Linear(state_dim + action_dim, hidden_dim)
        self.fc5 = nn.Linear(hidden_dim, hidden_dim)
        self.fc6 = nn.Linear(hidden_dim, 1)

    def forward(self, state, action):
        xu = torch.cat([state, action], 1)
        x1 = F.relu(self.fc1(xu))
        x1 = F.relu(self.fc2(x1))
        q1 = self.fc3(x1)

        x2 = F.relu(self.fc4(xu))
        x2 = F.relu(self.fc5(x2))
        q2 = self.fc6(x2)
        return q1, q2


# ==================== 经验回放池 ====================
class EnhancedReplayBuffer:
    def __init__(self, capacity=50000):
        self.buffer = deque(maxlen=capacity)
        self.experience_info = deque(maxlen=capacity)

    def push(self, state, action, reward, next_state, done, path_idx, similarity):
        self.buffer.append((state, action, reward, next_state, done))
        self.experience_info.append({'path_idx': path_idx, 'similarity': similarity})

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return None
        batch = random.sample(self.buffer, batch_size)
        state, action, reward, next_state, done = map(np.stack, zip(*batch))
        return (
            torch.FloatTensor(state).to(device),
            torch.FloatTensor(action).to(device),
            torch.FloatTensor(reward).unsqueeze(1).to(device),
            torch.FloatTensor(next_state).to(device),
            torch.FloatTensor(done).unsqueeze(1).to(device)
        )

    def get_top_k_per_path(self, num_paths, k=20):
        path_samples = {i: [] for i in range(num_paths)}
        for idx, info in enumerate(self.experience_info):
            path_idx = info['path_idx']
            similarity = info['similarity']
            path_samples[path_idx].append((idx, similarity, self.buffer[idx]))

        top_k_results = {}
        for path_idx in range(num_paths):
            samples = path_samples[path_idx]
            if len(samples) == 0:
                top_k_results[path_idx] = []
                continue
            samples.sort(key=lambda x: x[1], reverse=True)
            top_k = samples[:k]

            top_k_results[path_idx] = []
            for sample in top_k:
                normalized_state = sample[2][0]
                original_state = denormalize_state(normalized_state)
                # CPU核心数取整，内存和磁盘保留浮点
                original_state[0] = int(round(original_state[0]))
                triggered = execute_Tr(original_state)
                top_k_results[path_idx].append({
                    'state': original_state,
                    'similarity': sample[1],
                    'triggered': triggered
                })
        return top_k_results

    def __len__(self):
        return len(self.buffer)


# ==================== SAC Agent ====================
class SACAgent:
    def __init__(self, state_dim=3, action_dim=3):
        self.state_dim = state_dim
        self.action_dim = action_dim

        self.policy = GaussianPolicy(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.policy_optimizer = optim.Adam(self.policy.parameters(), lr=EXPERIMENT_CONFIG['ACTOR_LR'])

        self.critic = QNetwork(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_target = QNetwork(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_target.load_state_dict(self.critic.state_dict())
        self.critic_optimizer = optim.Adam(self.critic.parameters(), lr=EXPERIMENT_CONFIG['CRITIC_LR'])

        self.target_entropy = -action_dim
        self.log_alpha = torch.zeros(1, requires_grad=True, device=device)
        self.alpha_optimizer = optim.Adam([self.log_alpha], lr=EXPERIMENT_CONFIG['ALPHA_LR'])

        self.replay_buffer = EnhancedReplayBuffer()
        self.replay_train_count = 0

    def get_action(self, state, deterministic=False):
        normalized_state = normalize_state(state)
        state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)
        with torch.no_grad():
            if deterministic:
                _, _, action = self.policy.sample(state_tensor)
            else:
                action, _, _ = self.policy.sample(state_tensor)
        action = action.cpu().numpy()[0]
        return action

    def store_experience(self, state, action, reward, next_state, done, path_idx, similarity):
        normalized_state = normalize_state(state)
        normalized_next_state = normalize_state(next_state)
        self.replay_buffer.push(
            normalized_state, action, reward,
            normalized_next_state, done, path_idx, similarity
        )

    def replay_train(self):
        batch = self.replay_buffer.sample(EXPERIMENT_CONFIG['REPLAY_BATCH_SIZE'])
        if batch is None:
            return

        state, action, reward, next_state, done = batch
        with torch.no_grad():
            next_action, next_log_prob, _ = self.policy.sample(next_state)
            q1_next, q2_next = self.critic_target(next_state, next_action)
            q_next = torch.min(q1_next, q2_next)
            target_q = reward + (1 - done) * EXPERIMENT_CONFIG['GAMMA'] * (
                    q_next - self.log_alpha.exp() * next_log_prob)

        q1, q2 = self.critic(state, action)
        critic_loss = F.mse_loss(q1, target_q) + F.mse_loss(q2, target_q)
        self.critic_optimizer.zero_grad()
        critic_loss.backward()
        self.critic_optimizer.step()

        new_action, log_prob, _ = self.policy.sample(state)
        q1_new, q2_new = self.critic(state, new_action)
        q_new = torch.min(q1_new, q2_new)
        policy_loss = (self.log_alpha.exp() * log_prob - q_new).mean()

        self.policy_optimizer.zero_grad()
        policy_loss.backward()
        self.policy_optimizer.step()

        alpha_loss = -(self.log_alpha * (log_prob + self.target_entropy).detach()).mean()
        self.alpha_optimizer.zero_grad()
        alpha_loss.backward()
        self.alpha_optimizer.step()

        for param, target_param in zip(self.critic.parameters(), self.critic_target.parameters()):
            target_param.data.copy_(
                EXPERIMENT_CONFIG['TAU'] * param.data + (1 - EXPERIMENT_CONFIG['TAU']) * target_param.data)

        self.replay_train_count += 1
        if self.replay_train_count % 2 == 0:
            alpha_value = self.log_alpha.exp().item()
            print(f"  -> 回放训练次数 {self.replay_train_count}, Alpha={alpha_value:.4f}")


# ==================== 性能指标计算 ====================
def calculate_run_performance(run_idx, sac_results, training_time, total_steps, update_count, agent):
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)
    all_similarities = []
    total_samples = 0
    total_reward = 0

    for path_idx in range(num_paths):
        samples = sac_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']
            total_reward += reward
            all_similarities.append(similarity)
            total_samples += 1

    average_reward = total_reward / total_samples if total_samples > 0 else 0
    convergence = np.mean(all_similarities) if all_similarities else 0
    environment_adaptability = 1 / (np.std(all_similarities) + 1e-8) if len(all_similarities) > 1 else 0
    generalization_ability = convergence
    computational_efficiency = total_steps / training_time if training_time > 0 else 0
    policy_update_frequency = update_count / training_time if training_time > 0 else 0

    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0

    return {
        'Run': run_idx + 1,
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }


# ==================== Excel导出 ====================
def export_to_excel(all_sac_results, all_performance_data, target_paths, output_path="SAC_20_run.xlsx"):
    print("\n正在导出数据到 Excel...")
    all_sac_summary_data = []
    all_sac_detailed_data = []

    for run_idx, (sac_results, performance_data) in enumerate(zip(all_sac_results, all_performance_data)):
        # 路径摘要数据
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = sac_results[path_idx]
            if len(samples) == 0:
                all_sac_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue

            similarities = [s['similarity'] for s in samples]
            is_perfect = 'Yes' if any(abs(s - 1.0) < 0.001 for s in similarities) else 'No'
            all_sac_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })

            # 详细样本数据
            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']
                all_sac_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Sample ID': sample_idx + 1,
                    'CPU_Cores': int(state[0]),
                    'Memory_GB': round(state[1], 2),
                    'Disk_Space_GB': round(state[2], 2),
                    'Similarity': round(similarity, 4),
                    'Perfect Coverage': 'Yes' if abs(similarity - 1.0) < 0.001 else 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path))),
                    'Triggered Rules': ', '.join(map(str, sorted(triggered))),
                    'Intersection Count': len(target_path.intersection(triggered)),
                    'Target Rule Count': len(target_path)
                })

    sac_summary_df = pd.DataFrame(all_sac_summary_data)
    sac_detailed_df = pd.DataFrame(all_sac_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        sac_summary_df.to_excel(writer, sheet_name='SAC路径摘要', index=False)
        sac_detailed_df.to_excel(writer, sheet_name='SAC详细样本数据', index=False)

        selected_columns = [
            'Run', 'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df[selected_columns].to_excel(writer, sheet_name='性能指标', index=False)

        workbook = writer.book
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        perfect_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        # Sheet1 样式
        ws1 = writer.sheets['SAC路径摘要']
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row_idx in range(2, ws1.max_row + 1):
            if ws1.cell(row_idx, 9).value == 'Yes':
                for col_idx in range(1, ws1.max_column + 1):
                    ws1.cell(row_idx, col_idx).fill = perfect_fill
        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 18
        ws1.column_dimensions['D'].width = 15
        ws1.column_dimensions['E'].width = 18
        ws1.column_dimensions['F'].width = 15
        ws1.column_dimensions['G'].width = 15
        ws1.column_dimensions['H'].width = 15
        ws1.column_dimensions['I'].width = 18
        ws1.column_dimensions['J'].width = 50

        # Sheet2 样式
        ws2 = writer.sheets['SAC详细样本数据']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 12
        ws2.column_dimensions['E'].width = 12
        ws2.column_dimensions['F'].width = 16
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 18
        ws2.column_dimensions['I'].width = 40
        ws2.column_dimensions['J'].width = 40
        ws2.column_dimensions['K'].width = 18
        ws2.column_dimensions['L'].width = 18

        # Sheet3 样式
        ws3 = writer.sheets['性能指标']
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws3.column_dimensions[col].width = 20

    print(f"文件已成功保存到: {output_path}")
    print(f"  - Sheet1: SAC路径摘要 共计 {len(all_sac_summary_data)} 条记录")
    print(f"  - Sheet2: SAC详细样本数据 共计 {len(all_sac_detailed_data)} 条记录")
    print(f"  - Sheet3: 性能指标 共计 {len(all_performance_data)} 条记录")


# ==================== 训练流程 ====================
def train_sac_workflow():
    print("=" * 80)
    print("SAC训练开始 (资源配置规则)")
    print("奖励基于覆盖相似度")
    print("=" * 80)

    agent = SACAgent()
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n生成初始样本: 每路径 {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']} 个")
    path_samples = {}
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']

    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            # CPU核心数为整数，内存和磁盘为浮点数
            state = np.array([
                np.random.randint(min_vals[0], max_vals[0] + 1),
                np.random.uniform(min_vals[1], max_vals[1]),
                np.random.uniform(min_vals[2], max_vals[2])
            ], dtype=np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  路径 {path_idx + 1}/{num_paths}: 已生成 {len(samples)} 个样本")

    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size

    print(f"\n训练配置: 每批 {batch_size} 个样本, 每样本 {EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']} 步")
    print(f"总批次: {num_batches * num_paths} 批")
    print("-" * 80)

    for batch_idx in range(num_batches):
        print(f"\n批次 {batch_idx + 1}/{num_batches}")
        for path_idx in range(num_paths):
            target_path = target_paths[path_idx]
            batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]
            batch_rewards = []
            batch_similarities = []

            for sample_idx, initial_state in enumerate(batch_samples):
                state = initial_state.copy()
                episode_reward = 0
                final_similarity = 0

                for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                    action = agent.get_action(state, deterministic=False)
                    next_state = state + action
                    next_state = clip_state(next_state)
                    triggered = execute_Tr(next_state)
                    reward = unified_reward_function(triggered, target_path)
                    similarity = coverage_similarity(triggered, target_path)
                    done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)

                    agent.store_experience(
                        state, action, reward, next_state, done,
                        path_idx, similarity
                    )

                    state = next_state
                    episode_reward += reward
                    final_similarity = similarity
                    total_steps += 1

                batch_rewards.append(episode_reward)
                batch_similarities.append(final_similarity)

            avg_reward = np.mean(batch_rewards)
            avg_similarity = np.mean(batch_similarities)
            print(f"  路径 {path_idx + 1}: 平均奖励={avg_reward:.2f}, 平均相似度={avg_similarity:.4f}")

        print(f"\n  执行回放训练...")
        agent.replay_train()
        print(f"  经验池大小: {len(agent.replay_buffer)}")

    training_time = time.time() - start_time
    print("\n" + "=" * 80)
    print(f"SAC训练完成! 总耗时: {training_time:.2f} 秒, 总步数: {total_steps}")
    print(f"经验池大小: {len(agent.replay_buffer)}")
    print(f"回放训练次数: {agent.replay_train_count}")
    print("=" * 80)

    print(f"\n提取每路径 Top-{EXPERIMENT_CONFIG['TOP_K_SAMPLES']} 高相似度样本...")
    top_k_results = agent.replay_buffer.get_top_k_per_path(num_paths, EXPERIMENT_CONFIG['TOP_K_SAMPLES'])
    return agent, top_k_results, training_time, total_steps, agent.replay_train_count


# ==================== 主函数 ====================
def main():
    print("\n" + "=" * 80)
    print("SAC - 20次运行实验 (资源配置规则)")
    print("性能指标：覆盖相似度")
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    print(f"状态范围: cpu_cores[{min_vals[0]}-{max_vals[0]}], "
          f"memory_gb[{min_vals[1]}-{max_vals[1]}], "
          f"disk_space_gb[{min_vals[2]}-{max_vals[2]}]")
    print("=" * 80)

    all_sac_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'=' * 80}")
        print(f"开始第 {run_idx + 1}/{EXPERIMENT_CONFIG['NUM_RUNS']} 次运行")
        print(f"{'=' * 80}")

        sac_agent, sac_results, training_time, total_steps, update_count = train_sac_workflow()
        performance_data = calculate_run_performance(
            run_idx, sac_results, training_time, total_steps, update_count, sac_agent
        )

        all_sac_results.append(sac_results)
        all_performance_data.append(performance_data)

        print(f"\n第 {run_idx + 1} 次运行完成!")
        print(f"  总奖励: {performance_data['Total Reward']}")
        print(f"  平均奖励: {performance_data['Average Reward']}")
        print(f"  平均相似度: {performance_data['Average Similarity']}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"SAC_20_runs_resource_{timestamp}.xlsx"
    export_to_excel(all_sac_results, all_performance_data, target_paths, output_path)

    # 统计汇总
    print("\n" + "=" * 80)
    print("20次运行统计汇总")
    print("=" * 80)

    total_rewards = [p['Total Reward'] for p in all_performance_data]
    average_rewards = [p['Average Reward'] for p in all_performance_data]
    convergences = [p['Convergence'] for p in all_performance_data]
    environment_adaptabilities = [p['Environment Adaptability'] for p in all_performance_data]
    generalization_abilities = [p['Generalization Ability'] for p in all_performance_data]
    computational_efficiencies = [p['Computational Efficiency'] for p in all_performance_data]
    policy_update_frequencies = [p['Policy Update Frequency'] for p in all_performance_data]
    avg_similarities = [p['Average Similarity'] for p in all_performance_data]

    print(f"\n总奖励:")
    print(f"  平均值: {np.mean(total_rewards):.2f}")
    print(f"  标准差: {np.std(total_rewards):.2f}")

    print(f"\n平均奖励:")
    print(f"  平均值: {np.mean(average_rewards):.4f}")
    print(f"  标准差: {np.std(average_rewards):.4f}")

    print(f"\n收敛度 (平均相似度):")
    print(f"  平均值: {np.mean(convergences):.4f}")
    print(f"  标准差: {np.std(convergences):.4f}")

    print(f"\n环境适应性:")
    print(f"  平均值: {np.mean(environment_adaptabilities):.4f}")
    print(f"  标准差: {np.std(environment_adaptabilities):.4f}")

    print(f"\n泛化能力:")
    print(f"  平均值: {np.mean(generalization_abilities):.4f}")
    print(f"  标准差: {np.std(generalization_abilities):.4f}")

    print(f"\n计算效率 (步/秒):")
    print(f"  平均值: {np.mean(computational_efficiencies):.2f}")
    print(f"  标准差: {np.std(computational_efficiencies):.2f}")

    print(f"\n策略更新频率 (次/秒):")
    print(f"  平均值: {np.mean(policy_update_frequencies):.4f}")
    print(f"  标准差: {np.std(policy_update_frequencies):.4f}")

    print(f"\n平均相似度统计:")
    print(f"  平均值: {np.mean(avg_similarities):.4f}")
    print(f"  标准差: {np.std(avg_similarities):.4f}")

    print("\n" + "=" * 80)
    print(f"所有 {EXPERIMENT_CONFIG['NUM_RUNS']} 次运行完成!")
    print("=" * 80)


if __name__ == "__main__":
    main()