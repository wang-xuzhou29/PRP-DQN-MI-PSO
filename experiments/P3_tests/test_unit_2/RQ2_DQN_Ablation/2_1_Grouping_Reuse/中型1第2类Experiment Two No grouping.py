import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# ==================== 验证规则专用范围配置 ====================
CONFIG_DEPTH_MIN, CONFIG_DEPTH_MAX = 1, 7
PARAM_COUNT_MIN, PARAM_COUNT_MAX = 1, 180
SEC_WEIGHT_MIN, SEC_WEIGHT_MAX = 0.0, 1.0

# === 归一化/反归一化函数 ===
def normalize_state(state):
    """将状态归一化到 [0,1]"""
    cd, pc, sw = state
    cd_norm = (cd - CONFIG_DEPTH_MIN) / (CONFIG_DEPTH_MAX - CONFIG_DEPTH_MIN)
    pc_norm = (pc - PARAM_COUNT_MIN) / (PARAM_COUNT_MAX - PARAM_COUNT_MIN)
    sw_norm = (sw - SEC_WEIGHT_MIN) / (SEC_WEIGHT_MAX - SEC_WEIGHT_MIN)
    return (cd_norm, pc_norm, sw_norm)

def denormalize_state(state_norm):
    """将归一化状态还原：config_depth 和 param_count 取整，security_weight 保留两位小数"""
    cd_n, pc_n, sw_n = state_norm
    cd = int(round(cd_n * (CONFIG_DEPTH_MAX - CONFIG_DEPTH_MIN) + CONFIG_DEPTH_MIN))
    pc = int(round(pc_n * (PARAM_COUNT_MAX - PARAM_COUNT_MIN) + PARAM_COUNT_MIN))
    sw = sw_n * (SEC_WEIGHT_MAX - SEC_WEIGHT_MIN) + SEC_WEIGHT_MIN
    sw = np.clip(round(sw, 2), SEC_WEIGHT_MIN, SEC_WEIGHT_MAX)
    cd = np.clip(cd, CONFIG_DEPTH_MIN, CONFIG_DEPTH_MAX)
    pc = np.clip(pc, PARAM_COUNT_MIN, PARAM_COUNT_MAX)
    return (cd, pc, sw)

def normalize_value(value, min_val, max_val):
    return (value - min_val) / (max_val - min_val)

def denormalize_value(value_norm, min_val, max_val):
    return value_norm * (max_val - min_val) + min_val

# === 安全除法 ===
def safe_divide(numerator, denominator, default=0.0):
    if denominator == 0:
        return default
    return numerator / denominator

# === 奖励函数 ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    if prev_triggered is not None:
        prev_sim = jaccard_similarity(prev_triggered, target_path)
        improvement = sim - prev_sim
        reward += improvement * 5
    return reward

# === 执行验证规则函数（完整） ===
def execute_validation_rules(a):
    config_depth, param_count, security_weight = float(a[0]), int(a[1]), float(a[2])
    triggered = set()
    b = {}
    # 以下为原始验证规则（与您提供一致）
    if (config_depth >= 5) != (config_depth >= 5.8):
        b[0] = 1; triggered.add(1)
    if (config_depth >= 5) != (config_depth == 5):
        b[1] = 2; triggered.add(2)
    if config_depth >= 5:
        if (param_count >= 100 and security_weight >= 0.8) != (param_count != 100 and security_weight >= 0.8):
            b[2] = 3; triggered.add(3)
        if (param_count >= 100 and security_weight >= 0.8) != (param_count == 100 and security_weight >= 0.8):
            b[3] = 4; triggered.add(4)
        if (param_count >= 100 and security_weight >= 0.8) != (param_count >= 100 and security_weight != 0.8):
            b[4] = 5; triggered.add(5)
        if (param_count >= 100 and security_weight >= 0.8) != (param_count >= 100 and security_weight == 0.8):
            b[5] = 6; triggered.add(6)
        if (param_count >= 100 and security_weight >= 0.8) != (param_count >= 100 or security_weight >= 0.8):
            b[6] = 7; triggered.add(7)
        if param_count >= 100 and security_weight >= 0.8:
            if (security_weight >= 0.95) != (security_weight == 0.95):
                b[7] = 8; triggered.add(8)
            if (security_weight >= 0.95) != (security_weight != 0.95):
                b[8] = 9; triggered.add(9)
            if (security_weight >= 0.9) != (security_weight != 0.9):
                b[9] = 10; triggered.add(10)
            if (security_weight >= 0.9) != (security_weight == 0.9):
                b[10] = 11; triggered.add(11)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count != 50 and security_weight >= 0.6):
            b[11] = 12; triggered.add(12)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count == 50 and security_weight >= 0.6):
            b[12] = 13; triggered.add(13)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 60 and security_weight >= 0.6):
            b[13] = 14; triggered.add(14)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 50 and security_weight != 0.6):
            b[14] = 15; triggered.add(15)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 50 and security_weight == 0.6):
            b[15] = 16; triggered.add(16)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 53 and security_weight >= 0.6):
            b[16] = 17; triggered.add(17)
        elif param_count >= 50 and security_weight >= 0.6:
            if (security_weight >= 0.8) != (security_weight != 0.8):
                b[17] = 18; triggered.add(18)
            if (security_weight >= 0.8) != (security_weight == 0.8):
                b[18] = 19; triggered.add(19)
        if (param_count >= 25) != (param_count != 25):
            b[19] = 20; triggered.add(20)
        if (param_count >= 25) != (param_count >= 35):
            b[20] = 21; triggered.add(21)
    if (config_depth >= 3) != (config_depth != 3):
        b[21] = 22; triggered.add(22)
    if (config_depth >= 3) != (config_depth >= 3.9):
        b[22] = 23; triggered.add(23)
    elif config_depth >= 3:
        if (param_count >= 75 and security_weight >= 0.7) != (param_count >= 75 or security_weight >= 0.7):
            b[23] = 24; triggered.add(24)
        if (param_count >= 75 and security_weight >= 0.7) != (param_count == 75 and security_weight >= 0.7):
            b[24] = 25; triggered.add(25)
        if (param_count >= 75 and security_weight >= 0.7) != (param_count >= 75 and security_weight != 0.7):
            b[25] = 26; triggered.add(26)
        if (param_count >= 75 and security_weight >= 0.7) != (param_count >= 75 and security_weight == 0.7):
            b[26] = 27; triggered.add(27)
        if (param_count >= 75 and security_weight >= 0.7) != (param_count != 75 and security_weight >= 0.7):
            b[27] = 28; triggered.add(28)
        if param_count >= 75 and security_weight >= 0.7:
            if (security_weight >= 0.9) != (security_weight != 0.9):
                b[28] = 29; triggered.add(29)
            if (security_weight >= 0.9) != (security_weight == 0.9):
                b[29] = 30; triggered.add(30)
            if (security_weight >= 0.8) != (security_weight != 0.8):
                b[30] = 31; triggered.add(31)
            if (security_weight >= 0.8) != (security_weight == 0.8):
                b[31] = 32; triggered.add(32)
        if (param_count >= 40 and security_weight >= 0.5) != (param_count >= 40 or security_weight >= 0.5):
            b[32] = 33; triggered.add(33)
        if (param_count >= 40 and security_weight >= 0.5) != (param_count >= 40 and security_weight != 0.5):
            b[33] = 34; triggered.add(34)
        if (param_count >= 40 and security_weight >= 0.5) != (param_count >= 40 and security_weight == 0.5):
            b[34] = 35; triggered.add(35)
        if (param_count >= 40 and security_weight >= 0.5) != (param_count != 40 and security_weight >= 0.5):
            b[35] = 36; triggered.add(36)
        if (param_count >= 40 and security_weight >= 0.5) != (param_count == 40 and security_weight >= 0.5):
            b[36] = 37; triggered.add(37)
    if (config_depth >= 2) != (config_depth != 2):
        b[37] = 38; triggered.add(38)
    if (config_depth >= 2) != (config_depth >= 3):
        b[38] = 39; triggered.add(39)
    elif config_depth >= 2:
        if (param_count >= 50 and security_weight >= 0.6) != (param_count != 50 and security_weight >= 0.6):
            b[39] = 40; triggered.add(40)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count == 50 and security_weight >= 0.6):
            b[40] = 41; triggered.add(41)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 50 or security_weight >= 0.6):
            b[41] = 42; triggered.add(42)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 50 and security_weight == 0.6):
            b[42] = 43; triggered.add(43)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 50 and security_weight != 0.6):
            b[43] = 44; triggered.add(44)
        if (param_count >= 30) != (param_count != 30):
            b[44] = 45; triggered.add(45)
        if (param_count >= 30) != (param_count >= 45):
            b[45] = 46; triggered.add(46)
    if (param_count >= 150) != (param_count != 150):
        b[46] = 47; triggered.add(47)
    if (param_count >= 150) != (param_count >= 100):
        b[47] = 48; triggered.add(48)
    if param_count >= 150:
        if (config_depth >= 4 and security_weight >= 0.8) != (config_depth >= 4 or security_weight >= 0.8):
            b[48] = 49; triggered.add(49)
        if (config_depth >= 4 and security_weight >= 0.8) != (config_depth == 4 and security_weight >= 0.8):
            b[49] = 50; triggered.add(50)
        if (config_depth >= 4 and security_weight >= 0.8) != (config_depth != 4 and security_weight >= 0.8):
            b[50] = 51; triggered.add(51)
        if (config_depth >= 4 and security_weight >= 0.8) != (config_depth >= 4 and security_weight == 0.8):
            b[51] = 52; triggered.add(52)
        if (config_depth >= 4 and security_weight >= 0.8) != (config_depth >= 4 and security_weight != 0.8):
            b[52] = 53; triggered.add(53)
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth >= 3 and security_weight != 0.7):
            b[53] = 54; triggered.add(54)
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth >= 3 and security_weight == 0.7):
            b[54] = 55; triggered.add(55)
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth >= 3 or security_weight >= 0.7):
            b[55] = 56; triggered.add(56)
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth != 3 and security_weight >= 0.7):
            b[56] = 57; triggered.add(57)
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth == 3 and security_weight >= 0.7):
            b[57] = 58; triggered.add(58)
    if (param_count >= 100) != (param_count >= 125):
        b[58] = 59; triggered.add(59)
    if (param_count >= 100) != (param_count >= 200):
        b[59] = 60; triggered.add(60)
    elif param_count >= 100:
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth >= 5 and security_weight >= 0.81):
            b[60] = 61; triggered.add(61)
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth >= 5 and security_weight >= 0.7):
            b[61] = 62; triggered.add(62)
        if (security_weight >= 0.6) != (security_weight == 0.6):
            b[62] = 63; triggered.add(63)
        if (security_weight >= 0.6) != (security_weight != 0.6):
            b[63] = 64; triggered.add(64)
    if (param_count >= 50) != (param_count != 50):
        b[64] = 65; triggered.add(65)
    if (param_count >= 50) != (param_count >= 74):
        b[65] = 66; triggered.add(66)
    elif param_count >= 50:
        if (security_weight >= 0.8) != (security_weight != 0.8):
            b[66] = 67; triggered.add(67)
        if (security_weight >= 0.8) != (security_weight == 0.8):
            b[67] = 68; triggered.add(68)
        if (security_weight >= 0.6) != (security_weight != 0.6):
            b[68] = 69; triggered.add(69)
        if (security_weight >= 0.6) != (security_weight == 0.6):
            b[69] = 70; triggered.add(70)
    if (security_weight >= 0.9) != (security_weight != 0.9):
        b[70] = 71; triggered.add(71)
    if (security_weight >= 0.9) != (security_weight == 0.9):
        b[71] = 72; triggered.add(72)
    if security_weight >= 0.9:
        if (config_depth >= 4 and param_count >= 80) != (config_depth != 4 and param_count >= 80):
            b[72] = 73; triggered.add(73)
        if (config_depth >= 4 and param_count >= 80) != (config_depth == 4 and param_count >= 80):
            b[73] = 74; triggered.add(74)
        if (config_depth >= 4 and param_count >= 80) != (config_depth >= 4 and param_count != 80):
            b[74] = 75; triggered.add(75)
        if (config_depth >= 4 and param_count >= 80) != (config_depth >= 4 and param_count == 80):
            b[75] = 76; triggered.add(76)
        if (config_depth >= 4 and param_count >= 80) != (config_depth >= 4 and param_count >= 88):
            b[76] = 77; triggered.add(77)
        if (config_depth >= 4 and param_count >= 80) != (config_depth >= 4.4 and param_count >= 80):
            b[77] = 78; triggered.add(78)
        if (config_depth >= 3 and param_count >= 50) != (config_depth >= 3 and param_count != 50):
            b[78] = 79; triggered.add(79)
        if (config_depth >= 3 and param_count >= 50) != (config_depth >= 3 and param_count == 50):
            b[79] = 80; triggered.add(80)
        if (config_depth >= 3 and param_count >= 50) != (config_depth != 3 and param_count >= 50):
            b[80] = 81; triggered.add(81)
        if (config_depth >= 3 and param_count >= 50) != (config_depth == 3 and param_count >= 50):
            b[81] = 82; triggered.add(82)
        if (config_depth >= 3 and param_count >= 50) != (config_depth >= 3.6 and param_count >= 50):
            b[82] = 83; triggered.add(83)
        if (config_depth >= 3 and param_count >= 50) != (config_depth >= 3 and param_count >= 66):
            b[83] = 84; triggered.add(84)
    if (security_weight >= 0.7) != (security_weight != 0.7):
        b[84] = 85; triggered.add(85)
    if (security_weight >= 0.7) != (security_weight == 0.7):
        b[85] = 86; triggered.add(86)
    elif security_weight >= 0.7:
        if (config_depth >= 3 and param_count >= 60) != (config_depth >= 3 and param_count >= 66):
            b[86] = 87; triggered.add(87)
        if (config_depth >= 3 and param_count >= 60) != (config_depth >= 3 and param_count == 60):
            b[87] = 88; triggered.add(88)
        if (config_depth >= 3 and param_count >= 60) != (config_depth != 3 and param_count >= 60):
            b[88] = 89; triggered.add(89)
        if (config_depth >= 3 and param_count >= 60) != (config_depth == 3 and param_count >= 60):
            b[89] = 90; triggered.add(90)
        if (config_depth >= 3 and param_count >= 60) != (config_depth >= 3 and param_count >= 72):
            b[90] = 91; triggered.add(91)
        if (config_depth >= 3 and param_count >= 60) != (config_depth >= 3.3 and param_count >= 60):
            b[91] = 92; triggered.add(92)
        if (param_count >= 40) != (param_count != 40):
            b[92] = 93; triggered.add(93)
        if (param_count >= 40) != (param_count >= 25):
            b[93] = 94; triggered.add(94)
    if (security_weight >= 0.5) != (security_weight != 0.5):
        b[94] = 95; triggered.add(95)
    if (security_weight >= 0.5) != (security_weight == 0.5):
        b[95] = 96; triggered.add(96)
    elif security_weight >= 0.5:
        if (param_count >= 30) != (param_count >= 80):
            b[96] = 97; triggered.add(97)
        if (param_count >= 30) != (param_count >= 100):
            b[97] = 98; triggered.add(98)
    if (config_depth < 3) != (config_depth < 4.7):
        b[98] = 99; triggered.add(99)
    if (config_depth < 3) != (config_depth < 5):
        b[99] = 100; triggered.add(100)
    if (param_count < 50) != (param_count < 76):
        b[100] = 101; triggered.add(101)
    if (param_count < 50) != (param_count < 67):
        b[101] = 102; triggered.add(102)
    if (security_weight < 0.6) != (security_weight == 0.63):
        b[102] = 103; triggered.add(103)
    if (security_weight < 0.6) != (security_weight != 0.6):
        b[103] = 104; triggered.add(104)
    return triggered

# 设置别名，供后续使用
execute_Tr = execute_validation_rules

# === 目标路径组（验证规则，共10条） ===
targetPaths = [
    # A1
    {3, 7, 8, 11, 12, 14, 16, 17, 19, 22, 23, 24, 28, 30, 32, 35, 37, 40, 43, 47, 49, 51, 55, 57, 61, 62, 63, 65, 66, 68, 70, 72, 79, 81, 83, 84, 86, 96, 97, 98, 99, 100, 101, 102, 104},
    # A2
    {3, 7, 8, 11, 12, 14, 16, 17, 19, 24, 28, 30, 32, 35, 37, 40, 43, 47, 51, 52, 55, 58, 61, 62, 63, 65, 66, 68, 70, 72, 75, 79, 82, 84, 86, 96, 97, 98, 99, 100, 101, 102, 104},
    # A3
    {3, 6, 8, 11, 13, 16, 19, 22, 23, 25, 27, 30, 32, 35, 37, 41, 43, 47, 48, 49, 51, 55, 57, 59, 60, 61, 62, 63, 68, 70, 72, 73, 80, 81, 83, 86, 88, 89, 92, 96, 99, 100, 104},
    # A4
    {3, 7, 8, 11, 12, 19, 21, 24, 28, 30, 32, 33, 36, 40, 42, 45, 46, 47, 51, 52, 55, 58, 61, 62, 63, 65, 68, 70, 72, 75, 79, 86, 93, 94, 96, 97, 98, 99, 100, 104},
    # A5
    {1, 3, 7, 8, 11, 13, 16, 19, 24, 28, 30, 32, 35, 37, 41, 43, 47, 50, 52, 55, 58, 63, 66, 68, 70, 72, 75, 80, 82, 84, 86, 87, 90, 91, 96, 97, 98, 101, 102, 104},
    # A6
    {2, 3, 7, 8, 11, 12, 19, 20, 21, 24, 28, 30, 32, 33, 36, 40, 42, 45, 47, 50, 52, 55, 58, 63, 65, 68, 70, 72, 75, 79, 86, 93, 94, 96, 104},
    # A7
    {1, 9, 10, 15, 18, 24, 26, 29, 31, 33, 34, 42, 44, 47, 49, 53, 54, 56, 64, 67, 69, 71, 74, 75, 77, 80, 82, 85, 88, 90, 95, 98, 103},
    # A8
    {5, 7, 9, 10, 15, 18, 22, 24, 26, 29, 31, 33, 34, 38, 42, 44, 47, 60, 64, 67, 69, 71, 73, 81, 85, 89, 95, 103},
    # A9
    {4, 6, 8, 11, 13, 16, 19, 25, 27, 30, 32, 35, 37, 41, 43, 47, 51, 52, 55, 58, 60, 61, 62, 63, 68, 70, 72, 73, 76, 78, 80, 82, 86, 88, 90, 96, 99, 100, 104},
    # A10
    {3, 7, 8, 11, 13, 16, 19, 22, 27, 28, 30, 32, 35, 37, 38, 39, 41, 43, 47, 49, 51, 56, 57, 63, 68, 70, 72, 81, 86, 89, 96, 97, 98, 101, 104}
]
target_paths = [set(path) for path in targetPaths]
NUM_PATHS = len(targetPaths)

def jaccard_similarity(set1, set2):
    if not set2:
        return 0.0
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0

# === 路径相似度矩阵 ===
def compute_path_similarity_matrix(paths):
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix

def group_paths_by_similarity(paths, threshold_percentile=50):
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.percentile(avg_sim_scores, threshold_percentile)
    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]
    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)
    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group

# === 计算鲁棒性（邻域平均Jaccard相似度） ===
def compute_robustness(state, path, sample_size=9):
    base = execute_Tr(state[0], state[1], state[2])
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    # 针对验证规则的邻域步长（config_depth, param_count, security_weight）
    deltas = [
        (-1, -5, -0.05), (0, -5, 0), (1, -5, 0.05),
        (-1, 0, -0.05), (1, 0, 0.05),
        (-1, 5, -0.05), (0, 5, 0), (1, 5, 0.05),
        (0, 0, 0)
    ]

    for dc, dp, ds in deltas[:sample_size]:
        if dc == dp == ds == 0:
            continue
        neighbor_cd = int(np.clip(state[0] + dc, CONFIG_DEPTH_MIN, CONFIG_DEPTH_MAX))
        neighbor_pc = int(np.clip(state[1] + dp, PARAM_COUNT_MIN, PARAM_COUNT_MAX))
        neighbor_sw = np.clip(state[2] + ds, SEC_WEIGHT_MIN, SEC_WEIGHT_MAX)
        neighbor = (neighbor_cd, neighbor_pc, round(neighbor_sw, 2))
        n_trig = execute_Tr(neighbor[0], neighbor[1], neighbor[2])
        if not n_trig:
            continue
        rob += jaccard_similarity(n_trig, base)
        neighbors += 1

    return rob / neighbors if neighbors > 0 else 0.0

# === 生成样本（独立路径） ===
def generate_samples_for_all_paths(num_candidates=2000, top_k=200, run_id=1):
    BEST_WEIGHTS = [0.55, 0.25, 0.2]

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_individual.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Individual Path {path_id} (Weighted Screening) - Run {run_id}\n")
            f.write("config_depth param_count security_weight\tScore\tSimilarity\tRobustness\tLengthDiff\n")
            for s in samples:
                cd, pc, sw = s['state']
                f.write(
                    f"{cd} {pc} {sw:.2f}\t{s['score']:.4f}\t{s['similarity']:.4f}\t"
                    f"{s['robustness']:.4f}\t{s['length_diff']:.4f}\n"
                )

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_individual"

    for path_idx in range(len(targetPaths)):
        path = targetPaths[path_idx]
        candidate_samples = []
        attempts = 0
        max_attempts = num_candidates * 10

        while len(candidate_samples) < num_candidates and attempts < max_attempts:
            attempts += 1

            cd = random.randint(CONFIG_DEPTH_MIN, CONFIG_DEPTH_MAX)
            pc = random.randint(PARAM_COUNT_MIN, PARAM_COUNT_MAX)
            sw = round(random.uniform(SEC_WEIGHT_MIN, SEC_WEIGHT_MAX), 2)
            state = (cd, pc, sw)

            triggered = execute_Tr(cd, pc, sw)
            if not triggered:
                continue

            sim = jaccard_similarity(triggered, path)
            rob = compute_robustness(state, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))

            candidate_samples.append({
                'state': state,
                'similarity': sim,
                'robustness': rob,
                'length_diff': len_diff,
                'triggered': triggered
            })

        if candidate_samples:
            for sample in candidate_samples:
                score = (BEST_WEIGHTS[0] * sample['similarity'] +
                         BEST_WEIGHTS[1] * sample['robustness'] +
                         BEST_WEIGHTS[2] * sample['length_diff'])
                sample['score'] = score

            candidate_samples.sort(key=lambda x: x['score'], reverse=True)
            selected_samples = candidate_samples[:top_k]
            save_samples(path_id=path_idx + 1, samples=selected_samples, base_dir=base_dir)

# === 共享经验回放 ===
class SharedExperienceReplay:
    def __init__(self, capacity=20000):
        self.capacity = capacity
        self.buffer = deque(maxlen=self.capacity)
        self.priorities = deque(maxlen=self.capacity)

    def append(self, experience):
        self.buffer.append(experience)
        self.priorities.append(experience[-1])

    def sample(self, batch_size, alpha=0.6):
        if len(self.buffer) < batch_size:
            return [], [], []
        priorities = np.array(self.priorities, dtype=np.float64)
        priorities = np.power(priorities, alpha)
        probabilities = priorities / np.sum(priorities)
        batch_indices = np.random.choice(len(self.buffer), batch_size, p=probabilities, replace=False)
        batch = [self.buffer[idx] for idx in batch_indices]
        return batch, batch_indices, probabilities[batch_indices]

    def update_priorities(self, batch_indices, td_errors):
        for idx, td_error in zip(batch_indices, td_errors):
            if idx < len(self.priorities):
                self.priorities[idx] = max(abs(td_error), 1e-6)

    def __len__(self):
        return len(self.buffer)

    def get_high_reward_samples(self, target_path, num_samples=20):
        if len(self.buffer) == 0:
            return []
        samples_with_scores = []
        seen_states = set()
        for experience in self.buffer:
            state_tensor = experience[0]
            state_norm = state_tensor.cpu().numpy().flatten()
            state_tuple = denormalize_state((state_norm[0], state_norm[1], state_norm[2]))
            if state_tuple in seen_states:
                continue
            seen_states.add(state_tuple)
            triggered = execute_Tr(state_tuple[0], state_tuple[1], state_tuple[2])
            reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_scores.append((state_tuple, reward, sim, triggered))
        samples_with_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_scores[:num_samples]

def load_path_data(file_path):
    path_data = []
    if not os.path.exists(file_path):
        print(f"警告: 文件不存在 {file_path}")
        return path_data
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
            for line in lines[2:]:
                parts = line.strip().split("\t")
                if parts:
                    values = parts[0].split()
                    if len(values) >= 3:
                        cd = int(values[0])
                        pc = int(values[1])
                        sw = float(values[2])
                        path_data.append((cd, pc, sw))
    except Exception as e:
        print(f"读取文件 {file_path} 出错: {e}")
    return path_data

# === DQN 网络 ===
class DQN(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dims=[128, 64]):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, hidden_dims[0])
        self.fc2 = nn.Linear(hidden_dims[0], hidden_dims[1])
        self.fc3 = nn.Linear(hidden_dims[1], action_dim)
        self.dropout = nn.Dropout(0.1)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = self.dropout(x)
        x = torch.relu(self.fc2(x))
        return self.fc3(x)

# === DQN Agent with PER ===
class DQNAgentWithPER:
    def __init__(self, state_dim, action_dim, replay_buffer,
                 gamma=0.99, epsilon=1.0, epsilon_decay=0.995,
                 epsilon_min=0.1, learning_rate=0.001, alpha=0.6, beta=0.4):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer
        self.alpha = alpha
        self.beta = beta

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        """
        解码动作索引为三维步长（每个维度10种动作，共30维）
        针对验证规则设计的步长：
        - config_depth: 整数步长 [-3,-2,-1,0,1,2,3] 重复以填满10个（实际使用 [-3,-2,-1,0,0,1,2,3]）
        - param_count: 整数步长 [-20,-15,-10,-5,0,5,10,15,20,25]
        - security_weight: 浮点步长 [-0.2,-0.15,-0.1,-0.05,0,0.05,0.1,0.15,0.2,0.25]
        """
        # 为每个维度定义10个步长值
        step_lists = [
            [-3, -2, -1, 0, 0, 1, 2, 3, 3, 4],                # config_depth
            [-20, -15, -10, -5, 0, 5, 10, 15, 20, 25],        # param_count
            [-0.2, -0.15, -0.1, -0.05, 0, 0.05, 0.1, 0.15, 0.2, 0.25]  # security_weight
        ]

        dim = action_idx // 10
        idx = action_idx % 10
        if dim == 0:
            return (step_lists[0][idx], 0, 0)
        elif dim == 1:
            return (0, step_lists[1][idx], 0)
        else:
            return (0, 0, step_lists[2][idx])

    def get_legal_actions(self, state):
        legal_actions = []
        for action_idx in range(self.action_dim):
            dc, dp, ds = self.decode_action(action_idx)
            next_cd = state[0] + dc
            next_pc = state[1] + dp
            next_sw = state[2] + ds
            if (CONFIG_DEPTH_MIN <= next_cd <= CONFIG_DEPTH_MAX and
                PARAM_COUNT_MIN <= next_pc <= PARAM_COUNT_MAX and
                SEC_WEIGHT_MIN <= next_sw <= SEC_WEIGHT_MAX):
                legal_actions.append(action_idx)
        return legal_actions

    def act(self, state_norm, legal_actions=None):
        if legal_actions is None:
            legal_actions = list(range(self.action_dim))
        if not legal_actions:
            return None
        if random.random() < self.epsilon:
            return random.choice(legal_actions)
        state_tensor = torch.tensor(state_norm, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)[0]
        legal_q_values = q_values[legal_actions]
        best_legal_idx = torch.argmax(legal_q_values).item()
        return legal_actions[best_legal_idx]

    def store_transition(self, state, action, reward, next_state, done):
        state_norm = normalize_state(state)
        next_state_norm = normalize_state(next_state)
        state_tensor = torch.tensor(state_norm, dtype=torch.float32).unsqueeze(0).to(device)
        next_state_tensor = torch.tensor(next_state_norm, dtype=torch.float32).unsqueeze(0).to(device)

        with torch.no_grad():
            q_values = self.model(state_tensor)
            next_q_values = self.target_model(next_state_tensor)
            max_next_q = next_q_values.max(1)[0]
            target_q = reward + (self.gamma * max_next_q * (1 - done))
            td_error = abs(q_values[0][action].item() - target_q.item())

        self.replay_buffer.append((state_tensor, action, reward, next_state_tensor, done, td_error))
        return td_error

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return 0.0
        batch, batch_indices, probabilities = self.replay_buffer.sample(batch_size, alpha=self.alpha)
        if not batch:
            return 0.0

        states, actions, rewards, next_states, dones, _ = zip(*batch)
        weights = (len(self.replay_buffer) * probabilities) ** (-self.beta)
        weights = weights / weights.max()
        weights = torch.tensor(weights, dtype=torch.float32).to(device)

        states = torch.cat(states).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.cat(next_states).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q = self.target_model(next_states).max(1)[0].detach()
        target_q = rewards + (self.gamma * next_max_q * (1 - dones))

        td_errors = current_q - target_q
        weighted_loss = (td_errors.pow(2) * weights).mean()

        self.optimizer.zero_grad()
        weighted_loss.backward()
        torch.nn.utils.clip_grad_norm_(self.model.parameters(), 1.0)
        self.optimizer.step()

        new_priorities = abs(td_errors.detach().cpu().numpy())
        self.replay_buffer.update_priorities(batch_indices, new_priorities)

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay
        return weighted_loss.item()

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())

# === 训练（独立路径共享回放） ===
def generate_and_train_for_individual_paths(path_documents, repeats=5, batch_size=32, run_id=1):
    state_dim = 3
    action_dim = 30
    shared_replay_buffer = SharedExperienceReplay(capacity=20000)
    agent = DQNAgentWithPER(state_dim, action_dim, shared_replay_buffer)

    total_cumulative_reward = 0
    path_rewards = {}

    print(f"\n=== 第 {run_id}/20 次运行（独立路径训练）===")
    start_time = time.time()

    SAMPLES_PER_BATCH = 50
    NUM_BATCHES = 4
    STEPS_PER_SAMPLE = 3

    for path_idx in range(len(targetPaths)):
        path_id = path_idx + 1
        print(f"\n{'=' * 60}")
        print(f"Path {path_id}/{NUM_PATHS}")
        print(f"{'=' * 60}")

        file_path = os.path.join(path_documents, f"path{path_id}_individual.txt")
        if not os.path.exists(file_path):
            print(f"  警告: 文件不存在 {file_path}")
            continue

        path_data = load_path_data(file_path)
        if not path_data:
            print(f"  警告: Path {path_id} 无样本数据")
            continue

        target_path = targetPaths[path_idx]

        if path_idx not in path_rewards:
            path_rewards[path_idx] = 0

        for repeat_idx in range(repeats):
            print(f"\n  第 {repeat_idx + 1}/{repeats} 轮")

            for batch_idx in range(NUM_BATCHES):
                batch_start = batch_idx * SAMPLES_PER_BATCH
                batch_end = min(batch_start + SAMPLES_PER_BATCH, len(path_data))

                print(f"    第 {batch_idx + 1}/{NUM_BATCHES} 批 (样本 {batch_start}-{batch_end})")

                for sample_idx in range(batch_start, batch_end):
                    state = path_data[sample_idx]
                    prev_state = None
                    prev_triggered = None

                    for step in range(STEPS_PER_SAMPLE):
                        legal_actions = agent.get_legal_actions(state)
                        if not legal_actions:
                            break

                        state_norm = normalize_state(state)
                        action = agent.act(state_norm, legal_actions)
                        if action is None:
                            break

                        dc, dp, ds = agent.decode_action(action)
                        next_state = (
                            int(np.clip(state[0] + dc, CONFIG_DEPTH_MIN, CONFIG_DEPTH_MAX)),
                            int(np.clip(state[1] + dp, PARAM_COUNT_MIN, PARAM_COUNT_MAX)),
                            round(np.clip(state[2] + ds, SEC_WEIGHT_MIN, SEC_WEIGHT_MAX), 2)
                        )

                        triggered = execute_Tr(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered,
                                                prev_triggered, prev_state)
                        done = (step == STEPS_PER_SAMPLE - 1)

                        agent.store_transition(state, action, reward, next_state, done)

                        prev_state = state
                        prev_triggered = triggered
                        state = next_state

                        total_cumulative_reward += reward
                        path_rewards[path_idx] += reward

                if len(agent.replay_buffer) >= batch_size:
                    loss = agent.train(batch_size)
                    print(f"        第 {batch_idx+1} 批完成，损失: {loss:.4f}")

                if (batch_idx + 1) % 2 == 0:
                    agent.update_target_model()
                    print(f"        目标网络已更新 (第 {batch_idx+1} 批)")

        print(f"\nPath {path_id} 完成，总奖励: {path_rewards[path_idx]:.2f}")
        print(f"共享经验池大小: {len(shared_replay_buffer)}")

    training_time = time.time() - start_time
    print(f"\n=== 第 {run_id}/20 次运行完成，耗时: {training_time:.2f} 秒 ===")

    return agent, shared_replay_buffer, total_cumulative_reward, path_rewards, training_time

# === 生成Excel报告 ===
def create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir):
    os.makedirs(output_dir, exist_ok=True)

    similar_group_paths = [idx + 1 for idx in similar_group]
    isolated_group_paths = [idx + 1 for idx in isolated_group]

    wb = Workbook()

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    header_color = "4472C4"
    similar_group_color = "E2EFDA"
    isolated_group_color = "FCE4D6"
    stats_color = "FFF2CC"

    # === Sheet1: 路径相似度 ===
    ws_paths = wb.active
    ws_paths.title = "Path Similarities"

    path_headers = ['Path ID', 'Group'] + [f'Run {i}' for i in range(1, 21)] + \
                   ['Average', 'Max', 'Min', 'Std']

    for col, header in enumerate(path_headers, 1):
        cell = ws_paths.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_paths.row_dimensions[1].height = 30

    for path_id in range(1, NUM_PATHS + 1):
        row = path_id + 1
        if path_id in similar_group_paths:
            group_type = "High-correlation"
            row_color = similar_group_color
        elif path_id in isolated_group_paths:
            group_type = "Low-correlation"
            row_color = isolated_group_color
        else:
            group_type = "Ungrouped"
            row_color = "FFFFFF"

        cell = ws_paths.cell(row=row, column=1, value=f"Path {path_id}")
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        cell.border = thin_border

        cell = ws_paths.cell(row=row, column=2, value=group_type)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        cell.border = thin_border

        path_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            sim = run_data['path_similarities'].get(path_id, {}).get('avg_similarity', 0.0)
            path_similarities.append(sim)
            cell = ws_paths.cell(row=row, column=3 + run_idx, value=round(sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        stats_values = [
            round(np.mean(path_similarities), 4),
            round(np.max(path_similarities), 4),
            round(np.min(path_similarities), 4),
            round(np.std(path_similarities), 4)
        ]
        for i, value in enumerate(stats_values):
            cell = ws_paths.cell(row=row, column=23 + i, value=value)
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
            cell.font = Font(bold=True, size=10)
            cell.border = thin_border

    ws_paths.column_dimensions['A'].width = 13
    ws_paths.column_dimensions['B'].width = 16
    for col in range(3, 23):
        ws_paths.column_dimensions[get_column_letter(col)].width = 10
    for col in range(23, 27):
        ws_paths.column_dimensions[get_column_letter(col)].width = 13

    # === Sheet2: 分组统计 ===
    ws_groups = wb.create_sheet("Group Statistics")

    group_headers = ['Group Name', 'Included Paths'] + [f'Run {i}' for i in range(1, 21)] + ['Average', 'Std']

    for col, header in enumerate(group_headers, 1):
        cell = ws_groups.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_groups.row_dimensions[1].height = 30

    row = 2
    # 高相关组
    cell = ws_groups.cell(row=row, column=1, value="High-correlation group")
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    cell.border = thin_border

    cell = ws_groups.cell(row=row, column=2, value=','.join(map(str, similar_group_paths)))
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    cell.border = thin_border

    group_similarities = []
    for run_idx, run_data in enumerate(all_runs_data):
        group_sim = np.mean([
            run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
            for p in similar_group_paths
        ])
        group_similarities.append(group_sim)
        cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(group_sim, 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    cell = ws_groups.cell(row=row, column=23, value=round(np.mean(group_similarities), 4))
    cell.number_format = '0.0000'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border

    cell = ws_groups.cell(row=row, column=24, value=round(np.std(group_similarities), 4))
    cell.number_format = '0.0000'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border

    row += 1

    # 低相关组
    if isolated_group_paths:
        cell = ws_groups.cell(row=row, column=1, value="Low-correlation group")
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        cell.border = thin_border

        cell = ws_groups.cell(row=row, column=2, value=','.join(map(str, isolated_group_paths)))
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        cell.border = thin_border

        isolated_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            iso_sim = np.mean([
                run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
                for p in isolated_group_paths
            ])
            isolated_similarities.append(iso_sim)
            cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(iso_sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        cell = ws_groups.cell(row=row, column=23, value=round(np.mean(isolated_similarities), 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border

        cell = ws_groups.cell(row=row, column=24, value=round(np.std(isolated_similarities), 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border

    ws_groups.column_dimensions['A'].width = 16
    ws_groups.column_dimensions['B'].width = 22
    for col in range(3, 23):
        ws_groups.column_dimensions[get_column_letter(col)].width = 10
    ws_groups.column_dimensions[get_column_letter(23)].width = 14
    ws_groups.column_dimensions[get_column_letter(24)].width = 12

    # === Sheet3: 详细样本数据 ===
    ws_samples = wb.create_sheet("Detailed Samples")

    sample_headers = ['Run', 'Path ID', 'Sample ID', 'Config_Depth', 'Param_Count',
                      'Security_Weight', 'Similarity', 'Triggered Rules']

    for col, header in enumerate(sample_headers, 1):
        cell = ws_samples.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_samples.row_dimensions[1].height = 30

    sample_row = 2
    for run_idx, run_data in enumerate(all_runs_data, 1):
        for path_id in range(1, NUM_PATHS + 1):
            samples = run_data['path_samples'].get(path_id, [])
            if path_id in similar_group_paths:
                path_color = similar_group_color
            elif path_id in isolated_group_paths:
                path_color = isolated_group_color
            else:
                path_color = "FFFFFF"

            for sample_idx, (state_tuple, reward, sim, triggered) in enumerate(samples, 1):
                cd, pc, sw = state_tuple
                triggered_str = ','.join(map(str, sorted(triggered)))

                cell = ws_samples.cell(row=sample_row, column=1, value=f"Run {run_idx}")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=path_color, end_color=path_color, fill_type="solid")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=2, value=f"Path {path_id}")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=path_color, end_color=path_color, fill_type="solid")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=3, value=sample_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                for col_offset, value in enumerate([cd, pc, sw]):
                    cell = ws_samples.cell(row=sample_row, column=4 + col_offset, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=7, value=round(sim, 4))
                cell.number_format = '0.0000'
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=8, value=f"{{{triggered_str}}}")
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border

                sample_row += 1

    sample_widths = [13, 13, 11, 14, 14, 16, 12, 45]
    for i, width in enumerate(sample_widths, 1):
        ws_samples.column_dimensions[get_column_letter(i)].width = width

    # === Sheet4: 运行统计 ===
    ws_summary = wb.create_sheet("Run Statistics")

    summary_headers = ['Run', 'Time(s)', 'Avg Similarity', 'Max', 'Min',
                       'High Group Avg', 'Low Group Avg', 'Buffer Size']

    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_summary.row_dimensions[1].height = 30

    for run_idx, run_data in enumerate(all_runs_data, 1):
        row = run_idx + 1
        high_group_avg = np.mean([
            run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
            for p in similar_group_paths
        ])
        low_group_avg = 0.0
        if isolated_group_paths:
            low_group_avg = np.mean([
                run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
                for p in isolated_group_paths
            ])

        values = [
            f"Run {run_idx}",
            round(run_data['training_time'], 2),
            round(run_data['overall_avg_similarity'], 4),
            round(run_data['max_similarity'], 4),
            round(run_data['min_similarity'], 4),
            round(high_group_avg, 4),
            round(low_group_avg, 4),
            20000
        ]

        for col, value in enumerate(values, 1):
            cell = ws_summary.cell(row=row, column=col, value=value)
            if col == 1:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 2:
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col >= 3 and col <= 7:
                cell.number_format = '0.0000'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # 汇总行
    stat_row = len(all_runs_data) + 2
    stat_labels = ['', '', '', '', '', '', '', '']
    for col, label in enumerate(stat_labels, 1):
        cell = ws_summary.cell(row=stat_row, column=col, value=label)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    stat_row += 1
    training_times = [r['training_time'] for r in all_runs_data]
    overall_avgs = [r['overall_avg_similarity'] for r in all_runs_data]
    max_sims = [r['max_similarity'] for r in all_runs_data]
    min_sims = [r['min_similarity'] for r in all_runs_data]
    high_avgs = []
    low_avgs = []
    for run_data in all_runs_data:
        h_avg = np.mean([
            run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
            for p in similar_group_paths
        ])
        high_avgs.append(h_avg)
        if isolated_group_paths:
            l_avg = np.mean([
                run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0)
                for p in isolated_group_paths
            ])
            low_avgs.append(l_avg)

    stat_values = [
        '',
        round(np.sum(training_times), 2),
        round(np.mean(overall_avgs), 4),
        round(np.max(max_sims), 4),
        round(np.min(min_sims), 4),
        round(np.mean(high_avgs), 4),
        round(np.mean(low_avgs), 4) if low_avgs else 0.0,
        20000
    ]

    for col, value in enumerate(stat_values, 1):
        cell = ws_summary.cell(row=stat_row, column=col, value=value)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        if col == 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col == 2:
            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col >= 3 and col <= 7:
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    summary_widths = [13, 16, 18, 14, 14, 16, 16, 14]
    for i, width in enumerate(summary_widths, 1):
        ws_summary.column_dimensions[get_column_letter(i)].width = width

    output_path = os.path.join(output_dir, "20_runs_validation.xlsx")
    wb.save(output_path)
    print(f"\n综合Excel报告已生成: {output_path}")
    print(f"  包含4个工作表: 路径相似度、分组统计、详细样本、运行统计")

# === 主函数：20次运行 ===
def run_20_times_training():
    model_path_base = r"D:\Experiment\CNN\DQNNEW\saved_models_validation"
    path_documents = r"D:\Experiment\CNN\DQNNEW\path_samples_individual"
    output_dir = r"D:\Experiment\CNN\ComparisonExperiment2\excel_reports_validation"

    os.makedirs(model_path_base, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    similar_group, isolated_group = group_paths_by_similarity(targetPaths)
    similar_group_display = [idx + 1 for idx in similar_group]
    isolated_group_display = [idx + 1 for idx in isolated_group]

    print("=" * 60)
    print("20次运行 - 验证规则 (config_depth, param_count, security_weight)")
    print(f"状态范围: config_depth[{CONFIG_DEPTH_MIN},{CONFIG_DEPTH_MAX}], "
          f"param_count[{PARAM_COUNT_MIN},{PARAM_COUNT_MAX}], "
          f"security_weight[{SEC_WEIGHT_MIN},{SEC_WEIGHT_MAX}]")
    print(f"路径总数: {NUM_PATHS}")
    print(f"自动分组结果:")
    print(f"  高相关组: {similar_group_display}")
    print(f"  低相关组: {isolated_group_display}")
    print("动作空间: 30维 (每维度10种步长)")
    print("=" * 60)

    all_runs_data = []
    total_start_time = time.time()

    for run_id in range(1, 21):
        print(f"\n{'=' * 60}")
        print(f"开始第 {run_id}/20 次运行")
        print(f"{'=' * 60}")

        print(f"[Run {run_id}] 生成样本...")
        generate_samples_for_all_paths(num_candidates=2000, top_k=200, run_id=run_id)

        print(f"[Run {run_id}] 开始训练...")
        agent, shared_buffer, total_reward, path_rewards, training_time = \
            generate_and_train_for_individual_paths(path_documents, repeats=5,
                                                    batch_size=32, run_id=run_id)

        # 保存模型
        model_path = os.path.join(model_path_base, f"trained_model_run_{run_id}.pth")
        torch.save({
            'model_state_dict': agent.model.state_dict(),
            'optimizer_state_dict': agent.optimizer.state_dict(),
            'epsilon': agent.epsilon,
            'run_id': run_id,
            'ranges': {
                'config_depth': (CONFIG_DEPTH_MIN, CONFIG_DEPTH_MAX),
                'param_count': (PARAM_COUNT_MIN, PARAM_COUNT_MAX),
                'security_weight': (SEC_WEIGHT_MIN, SEC_WEIGHT_MAX)
            }
        }, model_path)
        print(f"[Run {run_id}] 模型已保存: {model_path}")

        # 收集数据
        run_data = {
            'run_id': run_id,
            'training_time': training_time,
            'total_reward': total_reward,
            'path_rewards': path_rewards,
            'path_similarities': {},
            'path_samples': {}
        }

        all_similarities = []
        for path_idx in range(len(targetPaths)):
            target_path = targetPaths[path_idx]
            high_reward_samples = shared_buffer.get_high_reward_samples(target_path, num_samples=20)

            if high_reward_samples:
                similarities = [sim for _, _, sim, _ in high_reward_samples]
                run_data['path_similarities'][path_idx + 1] = {
                    'avg_similarity': np.mean(similarities),
                    'max_similarity': np.max(similarities),
                    'min_similarity': np.min(similarities),
                    'sample_count': len(similarities)
                }
                run_data['path_samples'][path_idx + 1] = high_reward_samples
                all_similarities.extend(similarities)
            else:
                run_data['path_similarities'][path_idx + 1] = {
                    'avg_similarity': 0.0,
                    'max_similarity': 0.0,
                    'min_similarity': 0.0,
                    'sample_count': 0
                }
                run_data['path_samples'][path_idx + 1] = []

        if all_similarities:
            run_data['overall_avg_similarity'] = np.mean(all_similarities)
            run_data['max_similarity'] = np.max(all_similarities)
            run_data['min_similarity'] = np.min(all_similarities)
        else:
            run_data['overall_avg_similarity'] = 0.0
            run_data['max_similarity'] = 0.0
            run_data['min_similarity'] = 0.0

        all_runs_data.append(run_data)
        print(f"[Run {run_id}] 完成！总体平均相似度: {run_data['overall_avg_similarity']:.4f}")

    total_time = time.time() - total_start_time

    print("\n生成综合Excel报告...")
    create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir)

    print("\n" + "=" * 60)
    print("20次运行全部完成！")
    print(f"总耗时: {total_time:.2f} 秒 ({total_time/60:.2f} 分钟)")
    print(f"平均每次运行耗时: {total_time/20:.2f} 秒")
    print(f"\n平均相似度统计:")
    avg_sims = [r['overall_avg_similarity'] for r in all_runs_data]
    print(f"  总体平均: {np.mean(avg_sims):.4f}")
    print(f"  最大值: {np.max(avg_sims):.4f}")
    print(f"  最小值: {np.min(avg_sims):.4f}")
    print(f"  标准差: {np.std(avg_sims):.4f}")
    print(f"\n所有结果已保存至: {output_dir}")
    print("=" * 60)

if __name__ == "__main__":
    run_20_times_training()