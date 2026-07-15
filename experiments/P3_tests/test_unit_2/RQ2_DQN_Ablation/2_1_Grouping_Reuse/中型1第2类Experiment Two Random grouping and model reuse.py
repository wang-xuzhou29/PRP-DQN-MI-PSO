import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# ==================== 验证规则专用范围 ====================
STATE_RANGES = {
    'config_depth': (1, 7),      # 整数
    'param_count': (1, 180),     # 整数
    'security_weight': (0.0, 1.0)  # 浮点数，保留两位小数
}
STATE_NAMES = ('config_depth', 'param_count', 'security_weight')

# 为了兼容原代码中的 STATE_MIN / STATE_MAX（用于clip），转为numpy数组，注意类型为float
STATE_MIN = np.array([STATE_RANGES[name][0] for name in STATE_NAMES], dtype=np.float32)
STATE_MAX = np.array([STATE_RANGES[name][1] for name in STATE_NAMES], dtype=np.float32)

def clip_state(state):
    """裁剪状态到合法范围，并返回 tuple（第三个维度保留两位小数）"""
    state = np.array(state, dtype=np.float32)
    clipped = np.clip(state, STATE_MIN, STATE_MAX)
    # 前两个维度取整，第三个保留两位小数
    clipped[0] = int(round(clipped[0]))
    clipped[1] = int(round(clipped[1]))
    clipped[2] = round(clipped[2], 2)
    return tuple(clipped)

def random_state():
    """生成随机状态：前两个整数，第三个浮点数（两位小数）"""
    cd = random.randint(int(STATE_RANGES['config_depth'][0]), int(STATE_RANGES['config_depth'][1]))
    pc = random.randint(int(STATE_RANGES['param_count'][0]), int(STATE_RANGES['param_count'][1]))
    sw = round(random.uniform(STATE_RANGES['security_weight'][0], STATE_RANGES['security_weight'][1]), 2)
    return (cd, pc, sw)

class StateNormalizer:
    def __init__(self, ranges=None):
        self.ranges = ranges or STATE_RANGES
        self.names = STATE_NAMES

    def normalize(self, state):
        state = np.array(state, dtype=np.float32)
        normalized = np.zeros_like(state, dtype=np.float32)
        for i, name in enumerate(self.names):
            low, high = self.ranges[name]
            normalized[i] = (state[i] - low) / (high - low)
        return normalized

    def denormalize(self, normalized_state):
        normalized_state = np.array(normalized_state, dtype=np.float32)
        denormalized = np.zeros_like(normalized_state, dtype=np.float32)
        for i, name in enumerate(self.names):
            low, high = self.ranges[name]
            denormalized[i] = normalized_state[i] * (high - low) + low
        # 前两个取整，第三个保留两位小数
        denormalized[0] = int(round(np.clip(denormalized[0], low, high)))
        denormalized[1] = int(round(np.clip(denormalized[1], low, high)))
        denormalized[2] = round(np.clip(denormalized[2], low, high), 2)
        return denormalized

normalizer = StateNormalizer()

# === reward function ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward

# ==================== 执行验证规则函数 ====================
def execute_validation_rules(a):
    """
    参数a: 包含3个元素的元组或数组，对应 config_depth, param_count, security_weight
    返回: 触发的规则编号集合
    """
    config_depth, param_count, security_weight = float(a[0]), int(a[1]), float(a[2])
    triggered = set()
    b = {}

    # 以下为原始验证规则（与用户提供一致）
    if (config_depth >= 5) != (config_depth >= 5.8):
        b[0] = 1; triggered.add(1)
    if (config_depth >= 5) != (config_depth == 5):
        b[1] = 2; triggered.add(2)

    if config_depth >= 5:
        if (param_count >= 100 and security_weight >= 0.8) != (param_count != 100 and security_weight >= 0.8):
            b[2] = 3; triggered.add(3)
        if (param_count >= 100 and security_weight >= 0.8) != (param_count == 100 and security_weight >= 0.8):
            b[3] = 4; triggered.add(4)
        if (param_count >= 100 and security_weight >= 0.8) != (param_count >= 100 and security_weight != 0.8):
            b[4] = 5; triggered.add(5)
        if (param_count >= 100 and security_weight >= 0.8) != (param_count >= 100 and security_weight == 0.8):
            b[5] = 6; triggered.add(6)
        if (param_count >= 100 and security_weight >= 0.8) != (param_count >= 100 or security_weight >= 0.8):
            b[6] = 7; triggered.add(7)

        if param_count >= 100 and security_weight >= 0.8:
            if (security_weight >= 0.95) != (security_weight == 0.95):
                b[7] = 8; triggered.add(8)
            if (security_weight >= 0.95) != (security_weight != 0.95):
                b[8] = 9; triggered.add(9)
            if (security_weight >= 0.9) != (security_weight != 0.9):
                b[9] = 10; triggered.add(10)
            if (security_weight >= 0.9) != (security_weight == 0.9):
                b[10] = 11; triggered.add(11)

        if (param_count >= 50 and security_weight >= 0.6) != (param_count != 50 and security_weight >= 0.6):
            b[11] = 12; triggered.add(12)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count == 50 and security_weight >= 0.6):
            b[12] = 13; triggered.add(13)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 60 and security_weight >= 0.6):
            b[13] = 14; triggered.add(14)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 50 and security_weight != 0.6):
            b[14] = 15; triggered.add(15)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 50 and security_weight == 0.6):
            b[15] = 16; triggered.add(16)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 53 and security_weight >= 0.6):
            b[16] = 17; triggered.add(17)

        elif param_count >= 50 and security_weight >= 0.6:
            if (security_weight >= 0.8) != (security_weight != 0.8):
                b[17] = 18; triggered.add(18)
            if (security_weight >= 0.8) != (security_weight == 0.8):
                b[18] = 19; triggered.add(19)

        if (param_count >= 25) != (param_count != 25):
            b[19] = 20; triggered.add(20)
        if (param_count >= 25) != (param_count >= 35):
            b[20] = 21; triggered.add(21)

    if (config_depth >= 3) != (config_depth != 3):
        b[21] = 22; triggered.add(22)
    if (config_depth >= 3) != (config_depth >= 3.9):
        b[22] = 23; triggered.add(23)

    elif config_depth >= 3:
        if (param_count >= 75 and security_weight >= 0.7) != (param_count >= 75 or security_weight >= 0.7):
            b[23] = 24; triggered.add(24)
        if (param_count >= 75 and security_weight >= 0.7) != (param_count == 75 and security_weight >= 0.7):
            b[24] = 25; triggered.add(25)
        if (param_count >= 75 and security_weight >= 0.7) != (param_count >= 75 and security_weight != 0.7):
            b[25] = 26; triggered.add(26)
        if (param_count >= 75 and security_weight >= 0.7) != (param_count >= 75 and security_weight == 0.7):
            b[26] = 27; triggered.add(27)
        if (param_count >= 75 and security_weight >= 0.7) != (param_count != 75 and security_weight >= 0.7):
            b[27] = 28; triggered.add(28)

        if param_count >= 75 and security_weight >= 0.7:
            if (security_weight >= 0.9) != (security_weight != 0.9):
                b[28] = 29; triggered.add(29)
            if (security_weight >= 0.9) != (security_weight == 0.9):
                b[29] = 30; triggered.add(30)
            if (security_weight >= 0.8) != (security_weight != 0.8):
                b[30] = 31; triggered.add(31)
            if (security_weight >= 0.8) != (security_weight == 0.8):
                b[31] = 32; triggered.add(32)

        if (param_count >= 40 and security_weight >= 0.5) != (param_count >= 40 or security_weight >= 0.5):
            b[32] = 33; triggered.add(33)
        if (param_count >= 40 and security_weight >= 0.5) != (param_count >= 40 and security_weight != 0.5):
            b[33] = 34; triggered.add(34)
        if (param_count >= 40 and security_weight >= 0.5) != (param_count >= 40 and security_weight == 0.5):
            b[34] = 35; triggered.add(35)
        if (param_count >= 40 and security_weight >= 0.5) != (param_count != 40 and security_weight >= 0.5):
            b[35] = 36; triggered.add(36)
        if (param_count >= 40 and security_weight >= 0.5) != (param_count == 40 and security_weight >= 0.5):
            b[36] = 37; triggered.add(37)

    if (config_depth >= 2) != (config_depth != 2):
        b[37] = 38; triggered.add(38)
    if (config_depth >= 2) != (config_depth >= 3):
        b[38] = 39; triggered.add(39)

    elif config_depth >= 2:
        if (param_count >= 50 and security_weight >= 0.6) != (param_count != 50 and security_weight >= 0.6):
            b[39] = 40; triggered.add(40)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count == 50 and security_weight >= 0.6):
            b[40] = 41; triggered.add(41)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 50 or security_weight >= 0.6):
            b[41] = 42; triggered.add(42)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 50 and security_weight == 0.6):
            b[42] = 43; triggered.add(43)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 50 and security_weight != 0.6):
            b[43] = 44; triggered.add(44)

        if (param_count >= 30) != (param_count != 30):
            b[44] = 45; triggered.add(45)
        if (param_count >= 30) != (param_count >= 45):
            b[45] = 46; triggered.add(46)

    # 参数数量维度分析
    if (param_count >= 150) != (param_count != 150):
        b[46] = 47; triggered.add(47)
    if (param_count >= 150) != (param_count >= 100):
        b[47] = 48; triggered.add(48)

    if param_count >= 150:
        if (config_depth >= 4 and security_weight >= 0.8) != (config_depth >= 4 or security_weight >= 0.8):
            b[48] = 49; triggered.add(49)
        if (config_depth >= 4 and security_weight >= 0.8) != (config_depth == 4 and security_weight >= 0.8):
            b[49] = 50; triggered.add(50)
        if (config_depth >= 4 and security_weight >= 0.8) != (config_depth != 4 and security_weight >= 0.8):
            b[50] = 51; triggered.add(51)
        if (config_depth >= 4 and security_weight >= 0.8) != (config_depth >= 4 and security_weight == 0.8):
            b[51] = 52; triggered.add(52)
        if (config_depth >= 4 and security_weight >= 0.8) != (config_depth >= 4 and security_weight != 0.8):
            b[52] = 53; triggered.add(53)

        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth >= 3 and security_weight != 0.7):
            b[53] = 54; triggered.add(54)
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth >= 3 and security_weight == 0.7):
            b[54] = 55; triggered.add(55)
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth >= 3 or security_weight >= 0.7):
            b[55] = 56; triggered.add(56)
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth != 3 and security_weight >= 0.7):
            b[56] = 57; triggered.add(57)
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth == 3 and security_weight >= 0.7):
            b[57] = 58; triggered.add(58)

    if (param_count >= 100) != (param_count >= 125):
        b[58] = 59; triggered.add(59)
    if (param_count >= 100) != (param_count >= 200):
        b[59] = 60; triggered.add(60)

    elif param_count >= 100:
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth >= 5 and security_weight >= 0.81):
            b[60] = 61; triggered.add(61)
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth >= 5 and security_weight >= 0.7):
            b[61] = 62; triggered.add(62)
        if (security_weight >= 0.6) != (security_weight == 0.6):
            b[62] = 63; triggered.add(63)
        if (security_weight >= 0.6) != (security_weight != 0.6):
            b[63] = 64; triggered.add(64)

    if (param_count >= 50) != (param_count != 50):
        b[64] = 65; triggered.add(65)
    if (param_count >= 50) != (param_count >= 74):
        b[65] = 66; triggered.add(66)

    elif param_count >= 50:
        if (security_weight >= 0.8) != (security_weight != 0.8):
            b[66] = 67; triggered.add(67)
        if (security_weight >= 0.8) != (security_weight == 0.8):
            b[67] = 68; triggered.add(68)
        if (security_weight >= 0.6) != (security_weight != 0.6):
            b[68] = 69; triggered.add(69)
        if (security_weight >= 0.6) != (security_weight == 0.6):
            b[69] = 70; triggered.add(70)

    # 安全权重维度检查
    if (security_weight >= 0.9) != (security_weight != 0.9):
        b[70] = 71; triggered.add(71)
    if (security_weight >= 0.9) != (security_weight == 0.9):
        b[71] = 72; triggered.add(72)

    if security_weight >= 0.9:
        if (config_depth >= 4 and param_count >= 80) != (config_depth != 4 and param_count >= 80):
            b[72] = 73; triggered.add(73)
        if (config_depth >= 4 and param_count >= 80) != (config_depth == 4 and param_count >= 80):
            b[73] = 74; triggered.add(74)
        if (config_depth >= 4 and param_count >= 80) != (config_depth >= 4 and param_count != 80):
            b[74] = 75; triggered.add(75)
        if (config_depth >= 4 and param_count >= 80) != (config_depth >= 4 and param_count == 80):
            b[75] = 76; triggered.add(76)
        if (config_depth >= 4 and param_count >= 80) != (config_depth >= 4 and param_count >= 88):
            b[76] = 77; triggered.add(77)
        if (config_depth >= 4 and param_count >= 80) != (config_depth >= 4.4 and param_count >= 80):
            b[77] = 78; triggered.add(78)

        if (config_depth >= 3 and param_count >= 50) != (config_depth >= 3 and param_count != 50):
            b[78] = 79; triggered.add(79)
        if (config_depth >= 3 and param_count >= 50) != (config_depth >= 3 and param_count == 50):
            b[79] = 80; triggered.add(80)
        if (config_depth >= 3 and param_count >= 50) != (config_depth != 3 and param_count >= 50):
            b[80] = 81; triggered.add(81)
        if (config_depth >= 3 and param_count >= 50) != (config_depth == 3 and param_count >= 50):
            b[81] = 82; triggered.add(82)
        if (config_depth >= 3 and param_count >= 50) != (config_depth >= 3.6 and param_count >= 50):
            b[82] = 83; triggered.add(83)
        if (config_depth >= 3 and param_count >= 50) != (config_depth >= 3 and param_count >= 66):
            b[83] = 84; triggered.add(84)

    if (security_weight >= 0.7) != (security_weight != 0.7):
        b[84] = 85; triggered.add(85)
    if (security_weight >= 0.7) != (security_weight == 0.7):
        b[85] = 86; triggered.add(86)

    elif security_weight >= 0.7:
        if (config_depth >= 3 and param_count >= 60) != (config_depth >= 3 and param_count >= 66):
            b[86] = 87; triggered.add(87)
        if (config_depth >= 3 and param_count >= 60) != (config_depth >= 3 and param_count == 60):
            b[87] = 88; triggered.add(88)
        if (config_depth >= 3 and param_count >= 60) != (config_depth != 3 and param_count >= 60):
            b[88] = 89; triggered.add(89)
        if (config_depth >= 3 and param_count >= 60) != (config_depth == 3 and param_count >= 60):
            b[89] = 90; triggered.add(90)
        if (config_depth >= 3 and param_count >= 60) != (config_depth >= 3 and param_count >= 72):
            b[90] = 91; triggered.add(91)
        if (config_depth >= 3 and param_count >= 60) != (config_depth >= 3.3 and param_count >= 60):
            b[91] = 92; triggered.add(92)

        if (param_count >= 40) != (param_count != 40):
            b[92] = 93; triggered.add(93)
        if (param_count >= 40) != (param_count >= 25):
            b[93] = 94; triggered.add(94)

    if (security_weight >= 0.5) != (security_weight != 0.5):
        b[94] = 95; triggered.add(95)
    if (security_weight >= 0.5) != (security_weight == 0.5):
        b[95] = 96; triggered.add(96)

    elif security_weight >= 0.5:
        if (param_count >= 30) != (param_count >= 80):
            b[96] = 97; triggered.add(97)
        if (param_count >= 30) != (param_count >= 100):
            b[97] = 98; triggered.add(98)

    # Missing parameters detection based on dimensions
    if (config_depth < 3) != (config_depth < 4.7):
        b[98] = 99; triggered.add(99)
    if (config_depth < 3) != (config_depth < 5):
        b[99] = 100; triggered.add(100)
    if (param_count < 50) != (param_count < 76):
        b[100] = 101; triggered.add(101)
    if (param_count < 50) != (param_count < 67):
        b[101] = 102; triggered.add(102)
    if (security_weight < 0.6) != (security_weight == 0.63):
        b[102] = 103; triggered.add(103)
    if (security_weight < 0.6) != (security_weight != 0.6):
        b[103] = 104; triggered.add(104)

    return triggered

# 将规则函数赋给 execute_Tr，供后续调用
execute_Tr = execute_validation_rules

# ==================== 目标路径组（验证规则，10条） ====================
targetPaths = [
    # A1
    {3, 7, 8, 11, 12, 14, 16, 17, 19, 22, 23, 24, 28, 30, 32, 35, 37, 40, 43, 47, 49, 51, 55, 57, 61, 62, 63, 65, 66, 68, 70, 72, 79, 81, 83, 84, 86, 96, 97, 98, 99, 100, 101, 102, 104},
    # A2
    {3, 7, 8, 11, 12, 14, 16, 17, 19, 24, 28, 30, 32, 35, 37, 40, 43, 47, 51, 52, 55, 58, 61, 62, 63, 65, 66, 68, 70, 72, 75, 79, 82, 84, 86, 96, 97, 98, 99, 100, 101, 102, 104},
    # A3
    {3, 6, 8, 11, 13, 16, 19, 22, 23, 25, 27, 30, 32, 35, 37, 41, 43, 47, 48, 49, 51, 55, 57, 59, 60, 61, 62, 63, 68, 70, 72, 73, 80, 81, 83, 86, 88, 89, 92, 96, 99, 100, 104},
    # A4
    {3, 7, 8, 11, 12, 19, 21, 24, 28, 30, 32, 33, 36, 40, 42, 45, 46, 47, 51, 52, 55, 58, 61, 62, 63, 65, 68, 70, 72, 75, 79, 86, 93, 94, 96, 97, 98, 99, 100, 104},
    # A5
    {1, 3, 7, 8, 11, 13, 16, 19, 24, 28, 30, 32, 35, 37, 41, 43, 47, 50, 52, 55, 58, 63, 66, 68, 70, 72, 75, 80, 82, 84, 86, 87, 90, 91, 96, 97, 98, 101, 102, 104},
    # A6
    {2, 3, 7, 8, 11, 12, 19, 20, 21, 24, 28, 30, 32, 33, 36, 40, 42, 45, 47, 50, 52, 55, 58, 63, 65, 68, 70, 72, 75, 79, 86, 93, 94, 96, 104},
    # A7
    {1, 9, 10, 15, 18, 24, 26, 29, 31, 33, 34, 42, 44, 47, 49, 53, 54, 56, 64, 67, 69, 71, 74, 75, 77, 80, 82, 85, 88, 90, 95, 98, 103},
    # A8
    {5, 7, 9, 10, 15, 18, 22, 24, 26, 29, 31, 33, 34, 38, 42, 44, 47, 60, 64, 67, 69, 71, 73, 81, 85, 89, 95, 103},
    # A9
    {4, 6, 8, 11, 13, 16, 19, 25, 27, 30, 32, 35, 37, 41, 43, 47, 51, 52, 55, 58, 60, 61, 62, 63, 68, 70, 72, 73, 76, 78, 80, 82, 86, 88, 90, 96, 99, 100, 104},
    # A10
    {3, 7, 8, 11, 13, 16, 19, 22, 27, 28, 30, 32, 35, 37, 38, 39, 41, 43, 47, 49, 51, 56, 57, 63, 68, 70, 72, 81, 86, 89, 96, 97, 98, 101, 104}
]
target_paths = [set(path) for path in targetPaths]

def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0

def compute_path_similarity_matrix(paths):
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix

def group_paths_by_similarity(paths):
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.mean(avg_sim_scores)

    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]
    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)

    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group

# === random grouping ===
USE_KEYBOARD_INPUT_GROUP_SIZE = False
RANDOM_GROUP_SEED = None

def group_paths_randomly(paths, use_keyboard_input=False, seed=None):
    n_paths = len(paths)
    original_group1, original_group2 = group_paths_by_similarity(paths)
    default_group1_size = len(original_group1)

    group1_size = default_group1_size

    if use_keyboard_input:
        while True:
            user_input = input(
                f"Random group1Number of Paths,  1~{n_paths - 1}; "
                f" {default_group1_size}: "
            ).strip()

            if user_input == "":
                group1_size = default_group1_size
                break

            try:
                group1_size = int(user_input)
                if 1 <= group1_size <= n_paths - 1:
                    break
                print(f": Random group1 1~{n_paths - 1} .")
            except ValueError:
                print(": , .")

    if not (1 <= group1_size <= n_paths - 1):
        raise ValueError(f"Random group1 1~{n_paths - 1} ,  {group1_size}")

    all_indices = list(range(n_paths))
    rng = random.Random(seed) if seed is not None else random
    rng.shuffle(all_indices)

    random_group1 = sorted(all_indices[:group1_size])
    random_group2 = sorted(all_indices[group1_size:])

    return random_group1, random_group2, default_group1_size, len(original_group2)

def compute_robustness(state, path):
    """计算鲁棒性：邻域平均Jaccard相似度"""
    cd, pc, sw = state
    base = execute_Tr(cd, pc, sw)
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dc in [-1, 0, 1]:
        for dp in [-1, 0, 1]:
            for ds in [-0.05, 0, 0.05]:  # security_weight 步长较小
                if dc == dp == ds == 0:
                    continue
                neighbor = (cd + dc, pc + dp, sw + ds)
                neighbor = clip_state(neighbor)  # 确保在范围内
                n_trig = execute_Tr(neighbor[0], neighbor[1], neighbor[2])
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0

def compute_q_value_score(state, similar_model):
    """Q值分数：1 - 归一化Q值"""
    if similar_model is None:
        return 0.0

    try:
        normalized_state = normalizer.normalize(state)
        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = similar_model(state_tensor)
            max_q_value = torch.max(q_values).item()
            normalized_q = max_q_value / 20.0
            normalized_q = max(0.0, min(1.0, normalized_q))
            return 1.0 - normalized_q
    except:
        return 0.0

def generate_samples_for_similar_paths(similar_group, num_candidates=2000, top_k=200, run_id=1):
    SIMILAR_WEIGHTS = [0.55, 0.39, 0.06]

    def save_samples(path_id, samples, base_dir, group_type="similar"):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_{group_type}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"{group_type.title()} Group Path {path_id} - Run {run_id}\n")
            f.write("config_depth param_count security_weight\tScore\tSimilarity\tLengthDiff\tRobustness\n")
            for s in samples:
                cd, pc, sw = s['state']
                f.write(
                    f"{cd} {pc} {sw:.2f}\t{s['score']:.4f}\t{s['similarity']:.4f}\t{s['length_diff']:.4f}\t{s['robustness']:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_grouped"

    for path_idx in similar_group:
        path = target_paths[path_idx]
        path_id = path_idx + 1
        candidate_samples = []
        attempts = 0

        while len(candidate_samples) < num_candidates and attempts < num_candidates * 10:
            attempts += 1
            state = random_state()
            cd, pc, sw = state

            triggered = execute_Tr(cd, pc, sw)

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)

            candidate_samples.append({
                'state': state,
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            })

        if candidate_samples:
            for sample in candidate_samples:
                score = (SIMILAR_WEIGHTS[0] * sample['similarity'] +
                         SIMILAR_WEIGHTS[1] * sample['length_diff'] +
                         SIMILAR_WEIGHTS[2] * sample['robustness'])
                sample['score'] = score

            candidate_samples.sort(key=lambda x: x['score'], reverse=True)
            selected_samples = candidate_samples[:top_k]
            save_samples(path_id=path_id, samples=selected_samples, base_dir=base_dir, group_type="similar")

def generate_samples_for_isolated_paths(isolated_group, similar_model, num_candidates=2000, top_k=200, run_id=1):
    ISOLATED_WEIGHTS = [0.18, 0.21, 0.32, 0.29]

    def save_samples(path_id, samples, base_dir, group_type="isolated"):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_{group_type}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"{group_type.title()} Group Path {path_id} - Run {run_id}\n")
            f.write("config_depth param_count security_weight\tScore\tSimilarity\tLengthDiff\tRobustness\tQValueScore\n")
            for s in samples:
                cd, pc, sw = s['state']
                f.write(
                    f"{cd} {pc} {sw:.2f}\t{s['score']:.4f}\t{s['similarity']:.4f}\t{s['length_diff']:.4f}\t{s['robustness']:.4f}\t{s['q_value_score']:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_grouped"

    for path_idx in isolated_group:
        path = target_paths[path_idx]
        path_id = path_idx + 1
        candidate_samples = []
        attempts = 0

        while len(candidate_samples) < num_candidates and attempts < num_candidates * 10:
            attempts += 1
            state = random_state()
            cd, pc, sw = state

            triggered = execute_Tr(cd, pc, sw)

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            q_score = compute_q_value_score(state, similar_model)

            candidate_samples.append({
                'state': state,
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'q_value_score': q_score,
                'triggered': triggered
            })

        if candidate_samples:
            for sample in candidate_samples:
                score = (ISOLATED_WEIGHTS[0] * sample['similarity'] +
                         ISOLATED_WEIGHTS[1] * sample['length_diff'] +
                         ISOLATED_WEIGHTS[2] * sample['robustness'] +
                         ISOLATED_WEIGHTS[3] * sample['q_value_score'])
                sample['score'] = score

            candidate_samples.sort(key=lambda x: x['score'], reverse=True)
            selected_samples = candidate_samples[:top_k]
            save_samples(path_id=path_id, samples=selected_samples, base_dir=base_dir, group_type="isolated")

class GroupExperienceReplay:
    def __init__(self, capacity=20000):
        self.capacity = capacity
        self.buffer = deque(maxlen=self.capacity)
        self.priorities = deque(maxlen=self.capacity)
        self.sampled_indices = set()

    def append(self, experience):
        self.buffer.append(experience)
        self.priorities.append(experience[-1])

    def sample(self, batch_size, alpha=0.6):
        priorities = np.array(self.priorities) ** alpha
        probabilities = priorities / np.sum(priorities)
        batch_indices = np.random.choice(len(self.buffer), batch_size, p=probabilities)
        batch = [self.buffer[idx] for idx in batch_indices]
        return batch, batch_indices, probabilities[batch_indices]

    def update_priorities(self, batch_indices, td_errors):
        for idx, td_error in zip(batch_indices, td_errors):
            if idx < len(self.priorities):
                self.priorities[idx] = max(td_error, 1e-6)

    def __len__(self):
        return len(self.buffer)

    def get_high_reward_samples(self, target_path, num_samples=20):
        if len(self.buffer) == 0:
            return []

        samples_with_recalculated_scores = []
        for idx, experience in enumerate(self.buffer):
            if idx in self.sampled_indices:
                continue

            normalized_state_tensor = experience[0]
            normalized_state = normalized_state_tensor.cpu().numpy().flatten()
            state_tuple = tuple(normalizer.denormalize(normalized_state))

            cd, pc, sw = state_tuple
            triggered = execute_Tr(cd, pc, sw)
            new_reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_recalculated_scores.append((idx, state_tuple, new_reward, sim, triggered))

        samples_with_recalculated_scores.sort(key=lambda x: x[2], reverse=True)

        selected = samples_with_recalculated_scores[:num_samples]

        for item in selected:
            self.sampled_indices.add(item[0])

        return [(s[1], s[2], s[3], s[4]) for s in selected]

    def reset_sampled_indices(self):
        self.sampled_indices.clear()

def load_path_data(file_path):
    path_data = []
    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines[2:]:
            parts = line.strip().split("\t")
            values = parts[0].split()
            cd = int(values[0])
            pc = int(values[1])
            sw = float(values[2])
            path_data.append((cd, pc, sw))
    return path_data

class DQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, 128)
        self.fc2 = nn.Linear(128, 64)
        self.fc3 = nn.Linear(64, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)

class DQNAgentWithPER:
    def __init__(self, state_dim, action_dim, replay_buffer, gamma=0.99, epsilon=1.0, epsilon_decay=0.995,
                 epsilon_min=0.1, learning_rate=0.001, alpha=0.6, beta=0.4):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer
        self.alpha = alpha
        self.beta = beta

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        delta_values = [1, -1]
        dim = action_idx // 2
        delta_idx = action_idx % 2
        delta = delta_values[delta_idx]
        if dim == 0:
            return (delta, 0, 0)
        elif dim == 1:
            return (0, delta, 0)
        elif dim == 2:
            return (0, 0, delta)   # 步长 0.05 被映射为 +/-1，实际使用时乘 0.05

    def act(self, normalized_state):
        if random.random() < self.epsilon:
            return random.randrange(self.action_dim)
        state = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state)
        return torch.argmax(q_values, dim=1).item()

    def store_transition(self, normalized_state, action, reward, normalized_next_state, done):
        state = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        next_state = torch.tensor(normalized_next_state, dtype=torch.float32).unsqueeze(0).to(device)

        with torch.no_grad():
            q_values = self.model(state)
            next_q_values = self.target_model(next_state)
            max_next_q_values = next_q_values.max(1)[0]
            target_q_values = reward + (self.gamma * max_next_q_values * (1 - done))
            td_error = torch.abs(q_values[0][action] - target_q_values).item()

        self.replay_buffer.append((state, action, reward, next_state, done, td_error))
        return td_error

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return

        batch, batch_indices, probabilities = self.replay_buffer.sample(batch_size, alpha=self.alpha)
        states, actions, rewards, next_states, dones, _ = zip(*batch)

        weights = (len(self.replay_buffer) * probabilities) ** (-self.beta)
        weights = weights / weights.max()
        weights = torch.tensor(weights, dtype=torch.float32).to(device)

        states = torch.tensor(np.array([s.cpu().numpy().flatten() for s in states]), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array([ns.cpu().numpy().flatten() for ns in next_states]),
                                   dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))

        td_errors = current_q_values - target_q_values
        weighted_loss = (td_errors.pow(2) * weights).mean()

        self.optimizer.zero_grad()
        weighted_loss.backward()
        self.optimizer.step()

        new_priorities = torch.abs(td_errors).detach().cpu().numpy()
        self.replay_buffer.update_priorities(batch_indices, new_priorities)

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())

def train_group(group_paths, path_documents, replay_buffer, batch_size=32, group_name="", pretrained_model=None,
                sample_group_type=None):
    if sample_group_type is None:
        sample_group_type = 'similar' if group_name == '' else 'isolated'
    state_dim = 3
    action_dim = 6

    agent = DQNAgentWithPER(state_dim, action_dim, replay_buffer)

    if pretrained_model is not None:
        print(f"  {group_name}: ()...")
        agent.model.load_state_dict(pretrained_model.state_dict())
        agent.target_model.load_state_dict(pretrained_model.state_dict())
        print(f"  {group_name}: completed")

    path_rewards = {}

    print(f"Start training{group_name}, Included Paths: {[idx + 1 for idx in group_paths]}")
    start_time = time.time()

    BATCH_SIZE = 50
    N_SAMPLES = 200
    N_STEPS = 3
    N_ROUNDS = 5
    N_BATCHES = 4

    replay_count = 0

    for path_idx in group_paths:
        file_path = os.path.join(path_documents,
                                 f"path{path_idx + 1}_{sample_group_type}.txt")
        if not os.path.exists(file_path):
            print(f"    : Path {path_idx + 1}, ")
            continue

        path_data = load_path_data(file_path)
        target_path = target_paths[path_idx]

        if path_idx not in path_rewards:
            path_rewards[path_idx] = 0

        print(f"\n  Start training path  {path_idx + 1},  {N_ROUNDS} ")

        for round_idx in range(N_ROUNDS):
            print(f"    Path  {path_idx + 1} - Run  {round_idx + 1}/{N_ROUNDS} ")

            for batch_idx in range(N_BATCHES):
                batch_start = batch_idx * BATCH_SIZE
                batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)

                if batch_start >= len(path_data):
                    print(f"       {batch_idx + 1}: , ")
                    break

                print(f"       {batch_idx + 1}/{N_BATCHES} ( {batch_start}-{batch_end})")

                for sample_idx in range(batch_start, batch_end):
                    if sample_idx >= len(path_data):
                        break

                    state = path_data[sample_idx]
                    prev_state = None
                    prev_triggered = None

                    for step in range(N_STEPS):
                        normalized_state = normalizer.normalize(state)

                        legal_actions = []
                        for a in range(agent.action_dim):
                            dw, dt, dz = agent.decode_action(a)
                            # 第三个维度步长乘 0.05
                            cand_next = (state[0] + dw, state[1] + dt, state[2] + dz * 0.05)
                            cand_next = clip_state(cand_next)
                            # 判断是否合法（只要裁剪后等于原值即可，因为clip已经保证了范围）
                            legal_actions.append(a)

                        if not legal_actions:
                            break

                        if random.random() < agent.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        dw, dt, dz = agent.decode_action(action)
                        next_state = (state[0] + dw, state[1] + dt, state[2] + dz * 0.05)
                        next_state = clip_state(next_state)

                        normalized_next_state = normalizer.normalize(next_state)

                        cd, pc, sw = next_state
                        triggered = execute_Tr(cd, pc, sw)
                        reward = compute_reward(next_state, target_path, triggered,
                                                prev_triggered, prev_state)
                        done = (step == N_STEPS - 1)

                        agent.store_transition(normalized_state, action, reward, normalized_next_state, done)

                        prev_state = state
                        prev_triggered = triggered
                        state = next_state
                        path_rewards[path_idx] += reward

                if len(agent.replay_buffer) >= batch_size:
                    agent.train(batch_size)
                    replay_count += 1

                    if replay_count % 2 == 0:
                        agent.update_target_model()

            print(f"      Path  {path_idx + 1} - Run  {round_idx + 1} completed")

        print(f"  Path  {path_idx + 1}  {N_ROUNDS} All completed ")

    training_time = time.time() - start_time
    print(f"\n{group_name}completed!")
    print(f"  : Path completed{N_ROUNDS}()")
    print(f"  : {replay_count}")
    print(f"  : {training_time:.2f} seconds")
    print(f"  : {len(replay_buffer)}")

    return agent, path_rewards, training_time

def generate_and_train_grouped_paths_staged(path_documents, random_group1, random_group2, batch_size=32, run_id=1):
    print(f"\n===  {run_id}/20 (3 minutes, random grouping+model reuse) ===")
    random_group1_paths = [idx + 1 for idx in random_group1]
    random_group2_paths = [idx + 1 for idx in random_group2]

    print(f"Random group1Path (pretrained group): {random_group1_paths}")
    print(f"Random group2Path (model-reuse group): {random_group2_paths}")

    total_start_time = time.time()

    print(f"\n[1] Random group1...")
    generate_samples_for_similar_paths(random_group1, num_candidates=2000, top_k=200, run_id=run_id)

    print(f"\n[2] Random group1(, {5})...")
    group1_replay_buffer = GroupExperienceReplay(capacity=20000)
    group1_agent, group1_path_rewards, group1_training_time = train_group(
        random_group1, path_documents, group1_replay_buffer, batch_size,
        group_name="Random group1(pretrained group)", pretrained_model=None, sample_group_type="similar"
    )

    print(f"\n[3] Random group1Random group2...")
    generate_samples_for_isolated_paths(random_group2, group1_agent.model,
                                        num_candidates=2000, top_k=200, run_id=run_id)

    print(f"\n[4] Random group2(Random group1, {5})...")
    group2_replay_buffer = GroupExperienceReplay(capacity=20000)
    group2_agent, group2_path_rewards, group2_training_time = train_group(
        random_group2, path_documents, group2_replay_buffer, batch_size,
        group_name="Random group2(model-reuse group)", pretrained_model=group1_agent.model, sample_group_type="isolated"
    )

    total_path_rewards = {**group1_path_rewards, **group2_path_rewards}
    total_cumulative_reward = sum(total_path_rewards.values())
    total_training_time = time.time() - total_start_time

    print(f"\n===  {run_id}/20 completed, : {total_training_time:.2f} seconds ===")
    print(f"Random group1: {group1_training_time:.2f} seconds")
    print(f"Random group2: {group2_training_time:.2f} seconds")
    print(f" - Random group1: {len(group1_replay_buffer)}, Random group2: {len(group2_replay_buffer)}")

    return group1_agent, group2_agent, group1_replay_buffer, group2_replay_buffer, \
        total_cumulative_reward, total_path_rewards, total_training_time

def create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir):
    os.makedirs(output_dir, exist_ok=True)

    similar_group_paths = [idx + 1 for idx in similar_group]
    isolated_group_paths = [idx + 1 for idx in isolated_group]

    wb = Workbook()

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    header_color = "4472C4"
    similar_group_color = "E2EFDA"
    isolated_group_color = "FCE4D6"
    stats_color = "FFF2CC"

    ws_paths = wb.active
    ws_paths.title = "Path "

    path_headers = ['Path ID', ''] + [f'Run {i}' for i in range(1, 21)] + ['Average Similarity', 'Maximum Similarity',
                                                                                    'Minimum Similarity', 'Standard deviation']
    for col, header in enumerate(path_headers, 1):
        cell = ws_paths.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_paths.row_dimensions[1].height = 30

    for path_id in range(1, len(target_paths) + 1):
        row = path_id + 1

        if path_id in similar_group_paths:
            group_type = "Random group1(pretrained group)"
            row_color = similar_group_color
        elif path_id in isolated_group_paths:
            group_type = "Random group2(model-reuse group)"
            row_color = isolated_group_color
        else:
            group_type = "Ungrouped"
            row_color = "FFFFFF"

        cell = ws_paths.cell(row=row, column=1, value=f"Path {path_id}")
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        cell.border = thin_border

        cell = ws_paths.cell(row=row, column=2, value=group_type)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        cell.border = thin_border

        path_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            sim = run_data['path_similarities'].get(path_id, {}).get('avg_similarity', 0.0)
            path_similarities.append(sim)

            cell = ws_paths.cell(row=row, column=3 + run_idx, value=round(sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        stats_values = [
            round(np.mean(path_similarities), 4),
            round(np.max(path_similarities), 4),
            round(np.min(path_similarities), 4),
            round(np.std(path_similarities), 4)
        ]

        for i, value in enumerate(stats_values):
            cell = ws_paths.cell(row=row, column=23 + i, value=value)
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
            cell.font = Font(bold=True, size=10)
            cell.border = thin_border

    ws_paths.column_dimensions['A'].width = 13
    ws_paths.column_dimensions['B'].width = 16
    for col in range(3, 23):
        ws_paths.column_dimensions[get_column_letter(col)].width = 10
    for col in range(23, 27):
        ws_paths.column_dimensions[get_column_letter(col)].width = 13

    ws_groups = wb.create_sheet("")

    group_headers = ['Group Name', 'Included Paths'] + [f'Run {i}' for i in range(1, 21)] + ['Average Similarity', 'Standard deviation']
    for col, header in enumerate(group_headers, 1):
        cell = ws_groups.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_groups.row_dimensions[1].height = 30

    row = 2

    cell = ws_groups.cell(row=row, column=1, value="Random group1(pretrained group)")
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    cell.border = thin_border

    cell = ws_groups.cell(row=row, column=2, value=','.join(map(str, similar_group_paths)))
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    cell.border = thin_border

    group_similarities = []
    for run_idx, run_data in enumerate(all_runs_data):
        group_sim = np.mean(
            [run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0) for p in similar_group_paths])
        group_similarities.append(group_sim)

        cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(group_sim, 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    cell = ws_groups.cell(row=row, column=23, value=round(np.mean(group_similarities), 4))
    cell.number_format = '0.0000'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border

    cell = ws_groups.cell(row=row, column=24, value=round(np.std(group_similarities), 4))
    cell.number_format = '0.0000'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border

    row += 1

    if isolated_group_paths:
        cell = ws_groups.cell(row=row, column=1, value="Random group2(model-reuse group)")
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        cell.border = thin_border

        cell = ws_groups.cell(row=row, column=2, value=','.join(map(str, isolated_group_paths)))
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        cell.border = thin_border

        isolated_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            iso_sim = np.mean(
                [run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0) for p in isolated_group_paths])
            isolated_similarities.append(iso_sim)

            cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(iso_sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        cell = ws_groups.cell(row=row, column=23, value=round(np.mean(isolated_similarities), 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border

        cell = ws_groups.cell(row=row, column=24, value=round(np.std(isolated_similarities), 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border

    ws_groups.column_dimensions['A'].width = 16
    ws_groups.column_dimensions['B'].width = 22
    for col in range(3, 23):
        ws_groups.column_dimensions[get_column_letter(col)].width = 10
    ws_groups.column_dimensions[get_column_letter(23)].width = 14
    ws_groups.column_dimensions[get_column_letter(24)].width = 12

    ws_samples = wb.create_sheet("Detailed Sample Data")

    sample_headers = ['Run', 'Path ID', 'Sample ID', 'Config_Depth', 'Param_Count', 'Security_Weight', 'Similarity', 'Triggered Rule Set']
    for col, header in enumerate(sample_headers, 1):
        cell = ws_samples.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_samples.row_dimensions[1].height = 30

    sample_row = 2
    for run_idx, run_data in enumerate(all_runs_data, 1):
        for path_id in range(1, len(target_paths) + 1):
            samples = run_data['path_samples'].get(path_id, [])

            if path_id in similar_group_paths:
                path_color = similar_group_color
            elif path_id in isolated_group_paths:
                path_color = isolated_group_color
            else:
                path_color = "FFFFFF"

            for sample_idx, (state_tuple, reward, sim, triggered) in enumerate(samples, 1):
                cd, pc, sw = state_tuple
                triggered_str = ','.join(map(str, sorted(triggered)))

                cell = ws_samples.cell(row=sample_row, column=1, value=f"Run {run_idx}")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=path_color, end_color=path_color, fill_type="solid")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=2, value=f"Path {path_id}")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=path_color, end_color=path_color, fill_type="solid")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=3, value=sample_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                for col_offset, value in enumerate([cd, pc, sw]):
                    cell = ws_samples.cell(row=sample_row, column=4 + col_offset, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=7, value=round(sim, 4))
                cell.number_format = '0.0000'
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=8, value=f"{{{triggered_str}}}")
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border

                sample_row += 1

    sample_widths = [13, 13, 11, 14, 14, 16, 12, 45]
    for i, width in enumerate(sample_widths, 1):
        ws_samples.column_dimensions[get_column_letter(i)].width = width

    output_path = os.path.join(output_dir, "20_run_validation_random_reuse.xlsx")
    wb.save(output_path)
    print(f"\n Consolidated Excel report generated: {output_path}")

def run_20_times_training():
    model_path_base = r"D:\Experiment\CNN\DQNNEW\saved_models_validation_random_reuse"
    path_documents = r"D:\Experiment\CNN\DQNNEW\path_samples_grouped"
    output_dir = r"D:\Experiment\CNN\ComparisonExperiment2\excel_reports_validation_random_reuse"

    os.makedirs(model_path_base, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    similar_group, isolated_group, default_group1_size, default_group2_size = group_paths_randomly(
        target_paths,
        use_keyboard_input=USE_KEYBOARD_INPUT_GROUP_SIZE,
        seed=RANDOM_GROUP_SEED
    )

    similar_group_display = [idx + 1 for idx in similar_group]
    isolated_group_display = [idx + 1 for idx in isolated_group]

    print("=" * 60)
    print("20 - Validation rules + random grouping + model reuse")
    print(f"State ranges: config_depth[1,7], param_count[1,180], security_weight[0.0,1.0]")
    print("=" * 60)
    print("Training-scale configuration:")
    print("   Per path: 5 rounds")
    print("   Per round: 4 batches")
    print("   Per batch: 50 samples")
    print("   Per sample: 3 steps")
    print("   Sample generation: 2000 candidates -> 200 final samples")
    print("   Model saving: only parameters (optimized version)")
    print("   Grouping method: random grouping")
    print(f"   Default group size: Random group1={default_group1_size} paths, Random group2={default_group2_size} paths")
    print(f"   Keyboard input: {'enabled' if USE_KEYBOARD_INPUT_GROUP_SIZE else 'disabled'}")
    print(f"   Random seed: {RANDOM_GROUP_SEED if RANDOM_GROUP_SEED is not None else 'None (random per run)'}")
    print("=" * 60)
    print(f"\nAutomatic grouping results:")
    print(f"Random group1 (pretrained group): {similar_group_display}")
    print(f"Random group2 (model-reuse group): {isolated_group_display}")
    print("\n" + "=" * 60)

    all_runs_data = []
    total_start_time = time.time()

    for run_id in range(1, 21):
        print(f"\n{'=' * 60}")
        print(f"Start run  {run_id}/20")
        print(f"{'=' * 60}")

        group1_agent, group2_agent, group1_buffer, group2_buffer, total_cumulative_reward, path_rewards, training_time = \
            generate_and_train_grouped_paths_staged(path_documents, similar_group, isolated_group, batch_size=32,
                                                    run_id=run_id)

        group1_model_path = os.path.join(model_path_base, f"random_group1_model_run_{run_id}.pth")
        group2_model_path = os.path.join(model_path_base, f"random_group2_model_run_{run_id}.pth")

        torch.save(group1_agent.model.state_dict(), group1_model_path)
        torch.save(group2_agent.model.state_dict(), group2_model_path)

        print(f"[Run {run_id}] Model saved (optimized version)")

        group1_buffer.reset_sampled_indices()
        group2_buffer.reset_sampled_indices()

        run_data = {
            'run_id': run_id,
            'training_time': training_time,
            'total_reward': total_cumulative_reward,
            'path_rewards': path_rewards,
            'path_similarities': {},
            'path_samples': {}
        }

        all_similarities = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            path_id = path_idx + 1

            if path_id in similar_group_display:
                buffer = group1_buffer
            elif path_id in isolated_group_display:
                buffer = group2_buffer
            else:
                continue

            high_reward_samples = buffer.get_high_reward_samples(target_path, num_samples=20)

            if high_reward_samples:
                similarities = [sim for _, _, sim, _ in high_reward_samples]
                run_data['path_similarities'][path_idx + 1] = {
                    'avg_similarity': np.mean(similarities),
                    'max_similarity': np.max(similarities),
                    'min_similarity': np.min(similarities),
                    'sample_count': len(similarities)
                }
                run_data['path_samples'][path_idx + 1] = high_reward_samples
                all_similarities.extend(similarities)
            else:
                run_data['path_similarities'][path_idx + 1] = {
                    'avg_similarity': 0.0,
                    'max_similarity': 0.0,
                    'min_similarity': 0.0,
                    'sample_count': 0
                }
                run_data['path_samples'][path_idx + 1] = []

        if all_similarities:
            run_data['overall_avg_similarity'] = np.mean(all_similarities)
            run_data['max_similarity'] = np.max(all_similarities)
            run_data['min_similarity'] = np.min(all_similarities)
        else:
            run_data['overall_avg_similarity'] = 0.0
            run_data['max_similarity'] = 0.0
            run_data['min_similarity'] = 0.0

        all_runs_data.append(run_data)

        print(f"[Run {run_id}] completed! Overall Average Similarity: {run_data['overall_avg_similarity']:.4f}")
        print(f"{'=' * 60}\n")

    total_time = time.time() - total_start_time

    print("\nGenerating consolidated Excel report...")
    create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir)

    print("\n" + "=" * 60)
    print("20 runs all completed! - Validation rules + random grouping + model reuse")
    print("=" * 60)
    print(f"Summary:")
    print(f"  Per path: 5 rounds × 4 batches × 50 samples × 3 steps = 3000 steps/path")
    print(f"  Sample generation: 2000 candidates → 200 final samples")
    print(f"  Model saving: only parameters (optimized version)")
    print(f"  Total elapsed time: {total_time:.2f} seconds ({total_time / 60:.2f} minutes)")
    print(f"  Average elapsed time per run: {total_time / 20:.2f} seconds")
    print(f"\nAverage similarity statistics across runs:")
    avg_similarities = [r['overall_avg_similarity'] for r in all_runs_data]
    print(f"  Overall average: {np.mean(avg_similarities):.4f}")
    print(f"  Maximum: {np.max(avg_similarities):.4f}")
    print(f"  Minimum: {np.min(avg_similarities):.4f}")
    print(f"  Standard deviation: {np.std(avg_similarities):.4f}")
    print(f"\nAll results saved to: {output_dir}")
    print("=" * 60)

if __name__ == "__main__":
    run_20_times_training()