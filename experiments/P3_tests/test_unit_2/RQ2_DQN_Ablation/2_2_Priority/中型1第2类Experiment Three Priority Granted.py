import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# ==================== 验证规则专用范围 ====================
CONFIG_DEPTH_MIN, CONFIG_DEPTH_MAX = 1, 7
PARAM_COUNT_MIN, PARAM_COUNT_MAX = 1, 180
SEC_WEIGHT_MIN, SEC_WEIGHT_MAX = 0.0, 1.0

# === 归一化/反归一化（支持浮点数）===
def normalize_state(state):
    cd, pc, sw = state
    norm_cd = (cd - CONFIG_DEPTH_MIN) / (CONFIG_DEPTH_MAX - CONFIG_DEPTH_MIN)
    norm_pc = (pc - PARAM_COUNT_MIN) / (PARAM_COUNT_MAX - PARAM_COUNT_MIN)
    norm_sw = (sw - SEC_WEIGHT_MIN) / (SEC_WEIGHT_MAX - SEC_WEIGHT_MIN)
    return (norm_cd, norm_pc, norm_sw)

def denormalize_state(norm_state):
    n_cd, n_pc, n_sw = norm_state
    cd = int(round(n_cd * (CONFIG_DEPTH_MAX - CONFIG_DEPTH_MIN) + CONFIG_DEPTH_MIN))
    pc = int(round(n_pc * (PARAM_COUNT_MAX - PARAM_COUNT_MIN) + PARAM_COUNT_MIN))
    sw = n_sw * (SEC_WEIGHT_MAX - SEC_WEIGHT_MIN) + SEC_WEIGHT_MIN
    cd = max(CONFIG_DEPTH_MIN, min(CONFIG_DEPTH_MAX, cd))
    pc = max(PARAM_COUNT_MIN, min(PARAM_COUNT_MAX, pc))
    sw = max(SEC_WEIGHT_MIN, min(SEC_WEIGHT_MAX, round(sw, 2)))
    return (cd, pc, sw)

def is_valid_state(state):
    cd, pc, sw = state
    return (CONFIG_DEPTH_MIN <= cd <= CONFIG_DEPTH_MAX and
            PARAM_COUNT_MIN <= pc <= PARAM_COUNT_MAX and
            SEC_WEIGHT_MIN <= sw <= SEC_WEIGHT_MAX)

def clip_state(state):
    cd, pc, sw = state
    cd = max(CONFIG_DEPTH_MIN, min(CONFIG_DEPTH_MAX, int(round(cd))))
    pc = max(PARAM_COUNT_MIN, min(PARAM_COUNT_MAX, int(round(pc))))
    sw = max(SEC_WEIGHT_MIN, min(SEC_WEIGHT_MAX, round(sw, 2)))
    return (cd, pc, sw)

def random_state():
    cd = random.randint(CONFIG_DEPTH_MIN, CONFIG_DEPTH_MAX)
    pc = random.randint(PARAM_COUNT_MIN, PARAM_COUNT_MAX)
    sw = round(random.uniform(SEC_WEIGHT_MIN, SEC_WEIGHT_MAX), 2)
    return (cd, pc, sw)

# === Metric Collector（优先经验回放专用）===
class PrioritizedMetricsCollector:
    def __init__(self, experiment_name="Prioritized_DQN"):
        self.experiment_name = experiment_name
        self.start_time = None
        self.end_time = None
        self.total_reward = 0
        self.td_errors = []
        self.final_output_similarities = []
        self.action_improvements = []
        self.total_memory_usage = 0
        self.memory_check_count = 0
        self.step_count = 0
        self.episode_rewards = []
        self.episode_similarities = []
        self.episode_td_errors = []
        self.episode_epsilon_values = []
        self.episode_memory_usage = []
        self.priority_statistics = []
        self.importance_weights = []
        self.high_priority_samples_ratio = []
        self.priority_distribution_stats = []
        self.similar_paths_performance = []
        self.isolated_paths_performance = []
        self.milestone_data = {}
        self.convergence_window = 20
        self.convergence_threshold = 0.02
        self.convergence_detected_episode = None
        self.sample_efficiency_data = []
        self.performance_milestones = [0.6, 0.7, 0.75, 0.8]
        self.learning_curve_characteristics = {}
        self.early_vs_late_performance = {}

    def reset(self):
        self.__init__(self.experiment_name)

    def start_training(self):
        self.start_time = time.time()

    def end_training(self):
        self.end_time = time.time()

    def record_step_metrics(self, reward, td_error, triggered, target_path, priority=None, is_weight=None):
        self.step_count += 1
        self.total_reward += reward
        self.td_errors.append(td_error)
        if priority is not None:
            self.priority_statistics.append(priority)
        if is_weight is not None:
            self.importance_weights.append(is_weight)
        process = psutil.Process(os.getpid())
        current_memory = process.memory_info().rss / 1024 / 1024
        self.total_memory_usage += current_memory
        self.memory_check_count += 1

    def record_episode_metrics(self, episode, episode_reward, avg_similarity, avg_td_error, epsilon,
                               path_group="similar", priority_stats=None):
        self.episode_rewards.append(episode_reward)
        self.episode_similarities.append(avg_similarity)
        self.episode_td_errors.append(avg_td_error)
        self.episode_epsilon_values.append(epsilon)
        process = psutil.Process(os.getpid())
        current_memory = process.memory_info().rss / 1024 / 1024
        self.episode_memory_usage.append(current_memory)
        if priority_stats:
            self.priority_distribution_stats.append({
                'episode': episode,
                'mean_priority': priority_stats.get('mean_priority', 0),
                'max_priority': priority_stats.get('max_priority', 0),
                'min_priority': priority_stats.get('min_priority', 0),
                'high_priority_ratio': priority_stats.get('high_priority_ratio', 0)
            })
        if path_group == "similar":
            self.similar_paths_performance.append({
                'episode': episode,
                'reward': episode_reward,
                'similarity': avg_similarity,
                'td_error': avg_td_error
            })
        else:
            self.isolated_paths_performance.append({
                'episode': episode,
                'reward': episode_reward,
                'similarity': avg_similarity,
                'td_error': avg_td_error
            })
        if episode in [50, 100, 150, 200, 250, 300, 400, 450, 500]:
            self.milestone_data[episode] = {
                'avg_reward': np.mean(self.episode_rewards[-10:]) if len(self.episode_rewards) >= 10 else episode_reward,
                'avg_similarity': avg_similarity,
                'avg_td_error': avg_td_error,
                'epsilon': epsilon,
                'memory_usage': current_memory,
                'total_steps': self.step_count,
                'priority_stats': priority_stats
            }
        self._check_convergence(episode)
        self._check_performance_milestones(episode, avg_similarity)

    def _check_convergence(self, episode):
        if len(self.episode_similarities) >= self.convergence_window and self.convergence_detected_episode is None:
            recent = self.episode_similarities[-self.convergence_window:]
            if np.std(recent) < self.convergence_threshold:
                self.convergence_detected_episode = episode

    def _check_performance_milestones(self, episode, similarity):
        for milestone in self.performance_milestones:
            if similarity >= milestone and not any(data[1] == milestone for data in self.sample_efficiency_data):
                self.sample_efficiency_data.append((episode, milestone, self.step_count))

    def record_final_output_sample(self, triggered, target_path):
        if len(triggered | target_path) > 0:
            similarity = len(triggered & target_path) / len(triggered | target_path)
        else:
            similarity = 0.0
        self.final_output_similarities.append(similarity)

    def record_action_improvement(self, current_reward, prev_reward):
        if prev_reward is not None:
            improvement = current_reward - prev_reward
            self.action_improvements.append(1 if improvement > 0 else 0)

prioritized_metrics = PrioritizedMetricsCollector("Prioritized_DQN_Enhanced")

# === 奖励函数 ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward

# ==================== 执行验证规则函数 ====================
def execute_validation_rules(a):
    config_depth, param_count, security_weight = float(a[0]), int(a[1]), float(a[2])
    triggered = set()
    b = {}

    # 以下为原始验证规则（与用户提供一致）
    if (config_depth >= 5) != (config_depth >= 5.8):
        b[0] = 1; triggered.add(1)
    if (config_depth >= 5) != (config_depth == 5):
        b[1] = 2; triggered.add(2)

    if config_depth >= 5:
        if (param_count >= 100 and security_weight >= 0.8) != (param_count != 100 and security_weight >= 0.8):
            b[2] = 3; triggered.add(3)
        if (param_count >= 100 and security_weight >= 0.8) != (param_count == 100 and security_weight >= 0.8):
            b[3] = 4; triggered.add(4)
        if (param_count >= 100 and security_weight >= 0.8) != (param_count >= 100 and security_weight != 0.8):
            b[4] = 5; triggered.add(5)
        if (param_count >= 100 and security_weight >= 0.8) != (param_count >= 100 and security_weight == 0.8):
            b[5] = 6; triggered.add(6)
        if (param_count >= 100 and security_weight >= 0.8) != (param_count >= 100 or security_weight >= 0.8):
            b[6] = 7; triggered.add(7)

        if param_count >= 100 and security_weight >= 0.8:
            if (security_weight >= 0.95) != (security_weight == 0.95):
                b[7] = 8; triggered.add(8)
            if (security_weight >= 0.95) != (security_weight != 0.95):
                b[8] = 9; triggered.add(9)
            if (security_weight >= 0.9) != (security_weight != 0.9):
                b[9] = 10; triggered.add(10)
            if (security_weight >= 0.9) != (security_weight == 0.9):
                b[10] = 11; triggered.add(11)

        if (param_count >= 50 and security_weight >= 0.6) != (param_count != 50 and security_weight >= 0.6):
            b[11] = 12; triggered.add(12)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count == 50 and security_weight >= 0.6):
            b[12] = 13; triggered.add(13)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 60 and security_weight >= 0.6):
            b[13] = 14; triggered.add(14)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 50 and security_weight != 0.6):
            b[14] = 15; triggered.add(15)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 50 and security_weight == 0.6):
            b[15] = 16; triggered.add(16)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 53 and security_weight >= 0.6):
            b[16] = 17; triggered.add(17)

        elif param_count >= 50 and security_weight >= 0.6:
            if (security_weight >= 0.8) != (security_weight != 0.8):
                b[17] = 18; triggered.add(18)
            if (security_weight >= 0.8) != (security_weight == 0.8):
                b[18] = 19; triggered.add(19)

        if (param_count >= 25) != (param_count != 25):
            b[19] = 20; triggered.add(20)
        if (param_count >= 25) != (param_count >= 35):
            b[20] = 21; triggered.add(21)

    if (config_depth >= 3) != (config_depth != 3):
        b[21] = 22; triggered.add(22)
    if (config_depth >= 3) != (config_depth >= 3.9):
        b[22] = 23; triggered.add(23)

    elif config_depth >= 3:
        if (param_count >= 75 and security_weight >= 0.7) != (param_count >= 75 or security_weight >= 0.7):
            b[23] = 24; triggered.add(24)
        if (param_count >= 75 and security_weight >= 0.7) != (param_count == 75 and security_weight >= 0.7):
            b[24] = 25; triggered.add(25)
        if (param_count >= 75 and security_weight >= 0.7) != (param_count >= 75 and security_weight != 0.7):
            b[25] = 26; triggered.add(26)
        if (param_count >= 75 and security_weight >= 0.7) != (param_count >= 75 and security_weight == 0.7):
            b[26] = 27; triggered.add(27)
        if (param_count >= 75 and security_weight >= 0.7) != (param_count != 75 and security_weight >= 0.7):
            b[27] = 28; triggered.add(28)

        if param_count >= 75 and security_weight >= 0.7:
            if (security_weight >= 0.9) != (security_weight != 0.9):
                b[28] = 29; triggered.add(29)
            if (security_weight >= 0.9) != (security_weight == 0.9):
                b[29] = 30; triggered.add(30)
            if (security_weight >= 0.8) != (security_weight != 0.8):
                b[30] = 31; triggered.add(31)
            if (security_weight >= 0.8) != (security_weight == 0.8):
                b[31] = 32; triggered.add(32)

        if (param_count >= 40 and security_weight >= 0.5) != (param_count >= 40 or security_weight >= 0.5):
            b[32] = 33; triggered.add(33)
        if (param_count >= 40 and security_weight >= 0.5) != (param_count >= 40 and security_weight != 0.5):
            b[33] = 34; triggered.add(34)
        if (param_count >= 40 and security_weight >= 0.5) != (param_count >= 40 and security_weight == 0.5):
            b[34] = 35; triggered.add(35)
        if (param_count >= 40 and security_weight >= 0.5) != (param_count != 40 and security_weight >= 0.5):
            b[35] = 36; triggered.add(36)
        if (param_count >= 40 and security_weight >= 0.5) != (param_count == 40 and security_weight >= 0.5):
            b[36] = 37; triggered.add(37)

    if (config_depth >= 2) != (config_depth != 2):
        b[37] = 38; triggered.add(38)
    if (config_depth >= 2) != (config_depth >= 3):
        b[38] = 39; triggered.add(39)

    elif config_depth >= 2:
        if (param_count >= 50 and security_weight >= 0.6) != (param_count != 50 and security_weight >= 0.6):
            b[39] = 40; triggered.add(40)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count == 50 and security_weight >= 0.6):
            b[40] = 41; triggered.add(41)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 50 or security_weight >= 0.6):
            b[41] = 42; triggered.add(42)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 50 and security_weight == 0.6):
            b[42] = 43; triggered.add(43)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 50 and security_weight != 0.6):
            b[43] = 44; triggered.add(44)

        if (param_count >= 30) != (param_count != 30):
            b[44] = 45; triggered.add(45)
        if (param_count >= 30) != (param_count >= 45):
            b[45] = 46; triggered.add(46)

    # 参数数量维度分析
    if (param_count >= 150) != (param_count != 150):
        b[46] = 47; triggered.add(47)
    if (param_count >= 150) != (param_count >= 100):
        b[47] = 48; triggered.add(48)

    if param_count >= 150:
        if (config_depth >= 4 and security_weight >= 0.8) != (config_depth >= 4 or security_weight >= 0.8):
            b[48] = 49; triggered.add(49)
        if (config_depth >= 4 and security_weight >= 0.8) != (config_depth == 4 and security_weight >= 0.8):
            b[49] = 50; triggered.add(50)
        if (config_depth >= 4 and security_weight >= 0.8) != (config_depth != 4 and security_weight >= 0.8):
            b[50] = 51; triggered.add(51)
        if (config_depth >= 4 and security_weight >= 0.8) != (config_depth >= 4 and security_weight == 0.8):
            b[51] = 52; triggered.add(52)
        if (config_depth >= 4 and security_weight >= 0.8) != (config_depth >= 4 and security_weight != 0.8):
            b[52] = 53; triggered.add(53)

        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth >= 3 and security_weight != 0.7):
            b[53] = 54; triggered.add(54)
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth >= 3 and security_weight == 0.7):
            b[54] = 55; triggered.add(55)
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth >= 3 or security_weight >= 0.7):
            b[55] = 56; triggered.add(56)
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth != 3 and security_weight >= 0.7):
            b[56] = 57; triggered.add(57)
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth == 3 and security_weight >= 0.7):
            b[57] = 58; triggered.add(58)

    if (param_count >= 100) != (param_count >= 125):
        b[58] = 59; triggered.add(59)
    if (param_count >= 100) != (param_count >= 200):
        b[59] = 60; triggered.add(60)

    elif param_count >= 100:
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth >= 5 and security_weight >= 0.81):
            b[60] = 61; triggered.add(61)
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth >= 5 and security_weight >= 0.7):
            b[61] = 62; triggered.add(62)
        if (security_weight >= 0.6) != (security_weight == 0.6):
            b[62] = 63; triggered.add(63)
        if (security_weight >= 0.6) != (security_weight != 0.6):
            b[63] = 64; triggered.add(64)

    if (param_count >= 50) != (param_count != 50):
        b[64] = 65; triggered.add(65)
    if (param_count >= 50) != (param_count >= 74):
        b[65] = 66; triggered.add(66)

    elif param_count >= 50:
        if (security_weight >= 0.8) != (security_weight != 0.8):
            b[66] = 67; triggered.add(67)
        if (security_weight >= 0.8) != (security_weight == 0.8):
            b[67] = 68; triggered.add(68)
        if (security_weight >= 0.6) != (security_weight != 0.6):
            b[68] = 69; triggered.add(69)
        if (security_weight >= 0.6) != (security_weight == 0.6):
            b[69] = 70; triggered.add(70)

    # 安全权重维度检查
    if (security_weight >= 0.9) != (security_weight != 0.9):
        b[70] = 71; triggered.add(71)
    if (security_weight >= 0.9) != (security_weight == 0.9):
        b[71] = 72; triggered.add(72)

    if security_weight >= 0.9:
        if (config_depth >= 4 and param_count >= 80) != (config_depth != 4 and param_count >= 80):
            b[72] = 73; triggered.add(73)
        if (config_depth >= 4 and param_count >= 80) != (config_depth == 4 and param_count >= 80):
            b[73] = 74; triggered.add(74)
        if (config_depth >= 4 and param_count >= 80) != (config_depth >= 4 and param_count != 80):
            b[74] = 75; triggered.add(75)
        if (config_depth >= 4 and param_count >= 80) != (config_depth >= 4 and param_count == 80):
            b[75] = 76; triggered.add(76)
        if (config_depth >= 4 and param_count >= 80) != (config_depth >= 4 and param_count >= 88):
            b[76] = 77; triggered.add(77)
        if (config_depth >= 4 and param_count >= 80) != (config_depth >= 4.4 and param_count >= 80):
            b[77] = 78; triggered.add(78)

        if (config_depth >= 3 and param_count >= 50) != (config_depth >= 3 and param_count != 50):
            b[78] = 79; triggered.add(79)
        if (config_depth >= 3 and param_count >= 50) != (config_depth >= 3 and param_count == 50):
            b[79] = 80; triggered.add(80)
        if (config_depth >= 3 and param_count >= 50) != (config_depth != 3 and param_count >= 50):
            b[80] = 81; triggered.add(81)
        if (config_depth >= 3 and param_count >= 50) != (config_depth == 3 and param_count >= 50):
            b[81] = 82; triggered.add(82)
        if (config_depth >= 3 and param_count >= 50) != (config_depth >= 3.6 and param_count >= 50):
            b[82] = 83; triggered.add(83)
        if (config_depth >= 3 and param_count >= 50) != (config_depth >= 3 and param_count >= 66):
            b[83] = 84; triggered.add(84)

    if (security_weight >= 0.7) != (security_weight != 0.7):
        b[84] = 85; triggered.add(85)
    if (security_weight >= 0.7) != (security_weight == 0.7):
        b[85] = 86; triggered.add(86)

    elif security_weight >= 0.7:
        if (config_depth >= 3 and param_count >= 60) != (config_depth >= 3 and param_count >= 66):
            b[86] = 87; triggered.add(87)
        if (config_depth >= 3 and param_count >= 60) != (config_depth >= 3 and param_count == 60):
            b[87] = 88; triggered.add(88)
        if (config_depth >= 3 and param_count >= 60) != (config_depth != 3 and param_count >= 60):
            b[88] = 89; triggered.add(89)
        if (config_depth >= 3 and param_count >= 60) != (config_depth == 3 and param_count >= 60):
            b[89] = 90; triggered.add(90)
        if (config_depth >= 3 and param_count >= 60) != (config_depth >= 3 and param_count >= 72):
            b[90] = 91; triggered.add(91)
        if (config_depth >= 3 and param_count >= 60) != (config_depth >= 3.3 and param_count >= 60):
            b[91] = 92; triggered.add(92)

        if (param_count >= 40) != (param_count != 40):
            b[92] = 93; triggered.add(93)
        if (param_count >= 40) != (param_count >= 25):
            b[93] = 94; triggered.add(94)

    if (security_weight >= 0.5) != (security_weight != 0.5):
        b[94] = 95; triggered.add(95)
    if (security_weight >= 0.5) != (security_weight == 0.5):
        b[95] = 96; triggered.add(96)

    elif security_weight >= 0.5:
        if (param_count >= 30) != (param_count >= 80):
            b[96] = 97; triggered.add(97)
        if (param_count >= 30) != (param_count >= 100):
            b[97] = 98; triggered.add(98)

    # Missing parameters detection based on dimensions
    if (config_depth < 3) != (config_depth < 4.7):
        b[98] = 99; triggered.add(99)
    if (config_depth < 3) != (config_depth < 5):
        b[99] = 100; triggered.add(100)
    if (param_count < 50) != (param_count < 76):
        b[100] = 101; triggered.add(101)
    if (param_count < 50) != (param_count < 67):
        b[101] = 102; triggered.add(102)
    if (security_weight < 0.6) != (security_weight == 0.63):
        b[102] = 103; triggered.add(103)
    if (security_weight < 0.6) != (security_weight != 0.6):
        b[103] = 104; triggered.add(104)

    return triggered

# 设置 execute_Tr 指向验证规则函数
execute_Tr = execute_validation_rules

def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0

# ==================== 目标路径组（验证规则10条）====================
targetPaths = [
    # A1
    {3, 7, 8, 11, 12, 14, 16, 17, 19, 22, 23, 24, 28, 30, 32, 35, 37, 40, 43, 47, 49, 51, 55, 57, 61, 62, 63, 65, 66, 68, 70, 72, 79, 81, 83, 84, 86, 96, 97, 98, 99, 100, 101, 102, 104},
    # A2
    {3, 7, 8, 11, 12, 14, 16, 17, 19, 24, 28, 30, 32, 35, 37, 40, 43, 47, 51, 52, 55, 58, 61, 62, 63, 65, 66, 68, 70, 72, 75, 79, 82, 84, 86, 96, 97, 98, 99, 100, 101, 102, 104},
    # A3
    {3, 6, 8, 11, 13, 16, 19, 22, 23, 25, 27, 30, 32, 35, 37, 41, 43, 47, 48, 49, 51, 55, 57, 59, 60, 61, 62, 63, 68, 70, 72, 73, 80, 81, 83, 86, 88, 89, 92, 96, 99, 100, 104},
    # A4
    {3, 7, 8, 11, 12, 19, 21, 24, 28, 30, 32, 33, 36, 40, 42, 45, 46, 47, 51, 52, 55, 58, 61, 62, 63, 65, 68, 70, 72, 75, 79, 86, 93, 94, 96, 97, 98, 99, 100, 104},
    # A5
    {1, 3, 7, 8, 11, 13, 16, 19, 24, 28, 30, 32, 35, 37, 41, 43, 47, 50, 52, 55, 58, 63, 66, 68, 70, 72, 75, 80, 82, 84, 86, 87, 90, 91, 96, 97, 98, 101, 102, 104},
    # A6
    {2, 3, 7, 8, 11, 12, 19, 20, 21, 24, 28, 30, 32, 33, 36, 40, 42, 45, 47, 50, 52, 55, 58, 63, 65, 68, 70, 72, 75, 79, 86, 93, 94, 96, 104},
    # A7
    {1, 9, 10, 15, 18, 24, 26, 29, 31, 33, 34, 42, 44, 47, 49, 53, 54, 56, 64, 67, 69, 71, 74, 75, 77, 80, 82, 85, 88, 90, 95, 98, 103},
    # A8
    {5, 7, 9, 10, 15, 18, 22, 24, 26, 29, 31, 33, 34, 38, 42, 44, 47, 60, 64, 67, 69, 71, 73, 81, 85, 89, 95, 103},
    # A9
    {4, 6, 8, 11, 13, 16, 19, 25, 27, 30, 32, 35, 37, 41, 43, 47, 51, 52, 55, 58, 60, 61, 62, 63, 68, 70, 72, 73, 76, 78, 80, 82, 86, 88, 90, 96, 99, 100, 104},
    # A10
    {3, 7, 8, 11, 13, 16, 19, 22, 27, 28, 30, 32, 35, 37, 38, 39, 41, 43, 47, 49, 51, 56, 57, 63, 68, 70, 72, 81, 86, 89, 96, 97, 98, 101, 104}
]

def compute_path_similarity_matrix(paths):
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix

def group_paths_by_similarity(paths):
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.mean(avg_sim_scores)
    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]
    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)
    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group

# === 优先经验回放（使用状态去重）===
class PrioritizedExperienceReplay:
    def __init__(self, capacity=10000, alpha=0.6, beta_start=0.4, beta_frames=100000):
        self.capacity = capacity
        self.alpha = alpha
        self.beta_start = beta_start
        self.beta_frames = beta_frames
        self.frame = 1
        self.buffer = []
        self.priorities = np.zeros(capacity, dtype=np.float32)
        self.pos = 0
        self.size = 0
        self.max_priority = 1.0
        self.min_priority = 1.0

    def beta(self):
        return min(1.0, self.beta_start + (1.0 - self.beta_start) * self.frame / self.beta_frames)

    def append(self, experience):
        if len(self.buffer) < self.capacity:
            self.buffer.append(experience)
        else:
            self.buffer[self.pos] = experience
        self.priorities[self.pos] = self.max_priority
        self.pos = (self.pos + 1) % self.capacity
        self.size = min(self.size + 1, self.capacity)

    def sample(self, batch_size):
        if self.size < batch_size:
            return [], [], []
        priorities = self.priorities[:self.size]
        probs = priorities ** self.alpha
        probs /= probs.sum()
        indices = np.random.choice(self.size, batch_size, p=probs, replace=False)
        unique_batch = []
        unique_indices = []
        seen_states = set()
        for idx in indices:
            experience = self.buffer[idx]
            state_tensor = experience[0]
            state_tuple = tuple(state_tensor.cpu().numpy().flatten())
            if state_tuple not in seen_states:
                seen_states.add(state_tuple)
                unique_batch.append(experience)
                unique_indices.append(idx)
        if len(unique_batch) < batch_size:
            remaining_indices = [i for i in range(self.size) if i not in unique_indices]
            if remaining_indices:
                remaining_probs = priorities[remaining_indices] ** self.alpha
                remaining_probs /= remaining_probs.sum()
                needed = batch_size - len(unique_batch)
                additional_indices = np.random.choice(
                    remaining_indices,
                    min(needed, len(remaining_indices)),
                    p=remaining_probs,
                    replace=False
                )
                for idx in additional_indices:
                    experience = self.buffer[idx]
                    state_tensor = experience[0]
                    state_tuple = tuple(state_tensor.cpu().numpy().flatten())
                    if state_tuple not in seen_states:
                        seen_states.add(state_tuple)
                        unique_batch.append(experience)
                        unique_indices.append(idx)
        total = len(self.buffer)
        unique_indices = np.array(unique_indices)
        weights = (total * probs[unique_indices]) ** (-self.beta())
        weights /= weights.max()
        self.frame += 1
        return unique_batch, unique_indices, weights

    def update_priorities(self, indices, priorities):
        for idx, priority in zip(indices, priorities):
            if idx < self.size:
                self.priorities[idx] = priority
                self.max_priority = max(self.max_priority, priority)
                self.min_priority = min(self.min_priority, priority)

    def get_priority_statistics(self):
        if self.size == 0:
            return None
        priorities = self.priorities[:self.size]
        mean_priority = np.mean(priorities)
        max_priority = np.max(priorities)
        min_priority = np.min(priorities)
        high_priority_ratio = np.mean(priorities > mean_priority)
        return {
            'mean_priority': mean_priority,
            'max_priority': max_priority,
            'min_priority': min_priority,
            'high_priority_ratio': high_priority_ratio
        }

    def __len__(self):
        return self.size

    def get_high_reward_samples(self, target_path, num_samples=20):
        if len(self.buffer) == 0:
            return []
        samples_with_recalculated_scores = []
        for experience in self.buffer:
            state_tensor = experience[0]
            state_tuple = denormalize_state(state_tensor.cpu().numpy().flatten())
            triggered = execute_Tr(*state_tuple)
            new_reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_recalculated_scores.append((state_tuple, new_reward, sim, triggered))
        samples_with_recalculated_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_recalculated_scores[:num_samples]

def load_path_data(file_path):
    path_data = []
    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines[2:]:
            parts = line.strip().split("\t")
            values = parts[0].split()
            cd = int(values[0])
            pc = int(values[1])
            sw = float(values[2])
            path_data.append((cd, pc, sw))
    return path_data

# === DQN 网络 ===
class DQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, 128)
        self.fc2 = nn.Linear(128, 64)
        self.fc3 = nn.Linear(64, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)

# === 优先经验回放 DQN Agent ===
class PrioritizedDQNAgent:
    def __init__(self, state_dim, action_dim, replay_buffer, gamma=0.99, epsilon=1.0,
                 epsilon_decay=0.995, epsilon_min=0.1, learning_rate=0.001):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        # 针对验证规则设计步长（每个维度10种步长，共30个动作）
        step_lists = [
            [-3, -2, -1, 0, 1, 2, 3, 4, 5, 6],                        # config_depth
            [-30, -20, -10, -5, -2, 0, 2, 5, 10, 20],                 # param_count
            [-0.2, -0.15, -0.1, -0.05, -0.02, 0, 0.02, 0.05, 0.1, 0.2] # security_weight
        ]
        dim = action_idx // 10
        idx = action_idx % 10
        if dim == 0:
            return (step_lists[0][idx], 0, 0)
        elif dim == 1:
            return (0, step_lists[1][idx], 0)
        else:
            return (0, 0, step_lists[2][idx])

    def act(self, state):
        if random.random() < self.epsilon:
            return random.randrange(self.action_dim)
        norm_state = normalize_state(state)
        state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)
        return torch.argmax(q_values, dim=1).item()

    def store_transition(self, state, action, reward, next_state, done):
        norm_state = normalize_state(state)
        norm_next_state = normalize_state(next_state)
        state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
        next_state_tensor = torch.tensor(norm_next_state, dtype=torch.float32).unsqueeze(0).to(device)

        with torch.no_grad():
            q_values = self.model(state_tensor)
            next_q_values = self.target_model(next_state_tensor)
            max_next_q_values = next_q_values.max(1)[0]
            target_q_values = reward + (self.gamma * max_next_q_values * (1 - done))
            td_error = torch.abs(q_values[0][action] - target_q_values).item()

        self.replay_buffer.append((state_tensor, action, reward, next_state_tensor, done, td_error))
        return td_error

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return
        batch, batch_indices, is_weights = self.replay_buffer.sample(batch_size)
        if not batch:
            return

        states, actions, rewards, next_states, dones, _ = zip(*batch)

        states = torch.tensor(np.array([s.cpu().numpy().flatten() for s in states]), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array([ns.cpu().numpy().flatten() for ns in next_states]), dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)
        is_weights = torch.tensor(is_weights, dtype=torch.float32).to(device)

        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))

        td_errors = torch.abs(current_q_values - target_q_values)
        loss = (is_weights * (current_q_values - target_q_values) ** 2).mean()

        self.optimizer.zero_grad()
        loss.backward()
        self.optimizer.step()

        new_priorities = td_errors.detach().cpu().numpy() + 1e-6
        self.replay_buffer.update_priorities(batch_indices, new_priorities)

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())

# ==================== 样本生成（相似组）====================
def generate_samples_for_similar_paths(similar_group_indices, num_total=2000, top_k=200):
    def compute_robustness(state, path):
        base = execute_Tr(*state)
        if not base:
            return 0.0
        rob, neighbors = 0.0, 0
        # 邻域步长：config_depth ±1, param_count ±5, security_weight ±0.05
        for dc in [-1, 0, 1]:
            for dp in [-5, 0, 5]:
                for ds in [-0.05, 0, 0.05]:
                    if dc == dp == ds == 0:
                        continue
                    neighbor_state = (state[0] + dc, state[1] + dp, state[2] + ds)
                    if not is_valid_state(neighbor_state):
                        continue
                    neighbor = clip_state(neighbor_state)
                    n_trig = execute_Tr(*neighbor)
                    if not n_trig:
                        continue
                    rob += jaccard_similarity(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"prioritized_path{path_id}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Prioritized Path {path_id}\n")
            f.write("config_depth param_count security_weight\tScore\tSimilarity\tLengthDiff\tRobustness\n")
            for s in samples:
                cd, pc, sw = s[0]
                f.write(f"{cd} {pc} {sw:.2f}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_prioritized"
    for path_idx in similar_group_indices:
        path = targetPaths[path_idx]
        samples = []
        attempts = 0
        while len(samples) < top_k and attempts < num_total * 5:
            attempts += 1
            state = random_state()
            triggered = execute_Tr(*state)
            if not triggered:
                continue
            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            score = 0.55 * sim + 0.25 * len_diff + 0.2 * rob
            samples.append((state, score, sim, len_diff, rob))
        if samples:
            samples.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples[:top_k], base_dir=base_dir)

# ==================== 训练相似组（优先经验回放）====================
def prioritized_generate_and_train_for_similar_paths(agent, similar_group, path_documents, episodes=500,
                                                     batch_size=32, steps_per_test=3, replay_times=1,
                                                     is_isolated=False):
    trained_paths = set()
    global_replay_count = 0

    for episode in range(episodes):
        episode_reward = 0
        episode_similarities = []
        episode_td_errors = []

        for path_idx in similar_group:
            if path_idx in trained_paths:
                continue

            file_path = os.path.join(path_documents,
                                     f"prioritized_path{path_idx + 1}{'_isolated' if is_isolated else ''}.txt")
            path_data = load_path_data(file_path)
            target_path = targetPaths[path_idx]

            BATCH_SIZE = 50
            N_SAMPLES = 200
            N_STEPS = 3
            N_BATCHES = 4
            PATH_REPEAT = 5

            for repeat in range(PATH_REPEAT):
                for batch_idx in range(N_BATCHES):
                    batch_start = batch_idx * BATCH_SIZE
                    batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)

                    for test_data_idx in range(batch_start, batch_end):
                        if test_data_idx >= len(path_data):
                            break

                        state = path_data[test_data_idx]
                        prev_state = None
                        prev_triggered = None
                        prev_reward = None

                        for step in range(N_STEPS):
                            legal_actions = []
                            for a in range(agent.action_dim):
                                dc, dp, ds = agent.decode_action(a)
                                cand_next = (state[0] + dc, state[1] + dp, state[2] + ds)
                                if is_valid_state(cand_next):
                                    legal_actions.append(a)
                            if not legal_actions:
                                break

                            if random.random() < agent.epsilon:
                                action = random.choice(legal_actions)
                            else:
                                norm_state = normalize_state(state)
                                state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
                                with torch.no_grad():
                                    q_values = agent.model(state_tensor)[0]
                                action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                            dc, dp, ds = agent.decode_action(action)
                            next_state = clip_state((state[0] + dc, state[1] + dp, state[2] + ds))

                            triggered = execute_Tr(*next_state)
                            reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                            done = (step == N_STEPS - 1)

                            td_error = agent.store_transition(state, action, reward, next_state, done)

                            priority_stats = agent.replay_buffer.get_priority_statistics()
                            current_priority = priority_stats['mean_priority'] if priority_stats else 0

                            prioritized_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                    current_priority, None)

                            episode_similarities.append(jaccard_similarity(triggered, target_path))
                            episode_td_errors.append(td_error)

                            if prev_reward is not None:
                                prioritized_metrics.record_action_improvement(reward, prev_reward)

                            prev_state = state
                            prev_triggered = triggered
                            prev_reward = reward
                            state = next_state
                            episode_reward += reward

                    if len(agent.replay_buffer) >= batch_size:
                        agent.train(batch_size)
                        global_replay_count += 1
                        if global_replay_count % 2 == 0:
                            agent.update_target_model()

            trained_paths.add(path_idx)

        avg_similarity = np.mean(episode_similarities) if episode_similarities else 0
        avg_td_error = np.mean(episode_td_errors) if episode_td_errors else 0
        priority_stats = agent.replay_buffer.get_priority_statistics()

        prioritized_metrics.record_episode_metrics(episode, episode_reward, avg_similarity, avg_td_error,
                                                   agent.epsilon, "similar", priority_stats)

        if len(trained_paths) == len(similar_group):
            break

    return agent

# ==================== 生成孤立组样本（优先经验回放）====================
def generate_samples_for_isolated_paths_prioritized(agent_similar, isolated_group_indices, num_total=2000, top_k=200):
    def compute_q_value_normalized_complement(state, agent):
        norm_state = normalize_state(state)
        state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = agent.model(state_tensor)
        all_q_values = q_values[0].cpu().numpy()
        q_min = all_q_values.min()
        q_max = all_q_values.max()
        if q_max - q_min > 1e-6:
            normalized_q = (all_q_values.max() - q_min) / (q_max - q_min)
        else:
            normalized_q = 0.0
        return 1.0 - normalized_q

    def compute_robustness(state, path):
        base = execute_Tr(*state)
        if not base:
            return 0.0
        rob, neighbors = 0.0, 0
        for dc in [-1, 0, 1]:
            for dp in [-5, 0, 5]:
                for ds in [-0.05, 0, 0.05]:
                    if dc == dp == ds == 0:
                        continue
                    neighbor_state = (state[0] + dc, state[1] + dp, state[2] + ds)
                    if not is_valid_state(neighbor_state):
                        continue
                    neighbor = clip_state(neighbor_state)
                    n_trig = execute_Tr(*neighbor)
                    if not n_trig:
                        continue
                    rob += jaccard_similarity(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"prioritized_path{path_id}_isolated.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Prioritized Isolated Path {path_id}\n")
            f.write("config_depth param_count security_weight\tScore\tSimilarity\tLengthDiff\tRobustness\tQ_complement\n")
            for s in samples:
                cd, pc, sw = s[0]
                f.write(f"{cd} {pc} {sw:.2f}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\t{s[5]:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_prioritized"
    for path_idx in isolated_group_indices:
        path = targetPaths[path_idx]
        samples = []
        attempts = 0
        while len(samples) < top_k and attempts < num_total * 5:
            attempts += 1
            state = random_state()
            triggered = execute_Tr(*state)
            if not triggered:
                continue
            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            q_complement = compute_q_value_normalized_complement(state, agent_similar)
            score = 0.28 * sim + 0.1 * len_diff + 0.19 * rob + 0.43 * q_complement
            samples.append((state, score, sim, len_diff, rob, q_complement))
        if samples:
            samples.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples[:top_k], base_dir=base_dir)

# ==================== 训练孤立组（两阶段优先经验回放）====================
def prioritized_generate_and_train_for_isolated_paths(agent_similar, agent_isolated, similar_group,
                                                      isolated_group, path_documents, episodes=500, batch_size=32,
                                                      is_isolated=True):
    trained_paths = set()
    global_replay_count = 0

    stage1_samples_pool = {}
    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_similar.replay_buffer.get_high_reward_samples(target_path, num_samples=100)
        stage1_samples_pool[path_idx] = high_reward_samples

    for episode in range(episodes):
        episode_reward = 0
        episode_similarities = []
        episode_td_errors = []

        for path_idx in isolated_group:
            if path_idx in trained_paths:
                continue

            file_path = os.path.join(path_documents, f"prioritized_path{path_idx + 1}_isolated.txt")
            stage2_path_data = load_path_data(file_path)
            stage1_samples = stage1_samples_pool.get(path_idx, [])
            target_path = targetPaths[path_idx]

            BATCH_SIZE = 50
            N_SAMPLES_STAGE2 = min(140, len(stage2_path_data))
            N_SAMPLES_STAGE1 = min(60, len(stage1_samples))
            N_STEPS = 3
            N_BATCHES_STAGE2 = (N_SAMPLES_STAGE2 + BATCH_SIZE - 1) // BATCH_SIZE
            N_BATCHES_STAGE1 = (N_SAMPLES_STAGE1 + BATCH_SIZE - 1) // BATCH_SIZE
            PATH_REPEAT = 5

            for repeat in range(PATH_REPEAT):
                # Stage2
                for batch_idx in range(N_BATCHES_STAGE2):
                    batch_start = batch_idx * BATCH_SIZE
                    batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES_STAGE2)

                    for test_data_idx in range(batch_start, batch_end):
                        if test_data_idx >= len(stage2_path_data):
                            break

                        state = stage2_path_data[test_data_idx]
                        prev_state = None
                        prev_triggered = None
                        prev_reward = None

                        for step in range(N_STEPS):
                            legal_actions = []
                            for a in range(agent_isolated.action_dim):
                                dc, dp, ds = agent_isolated.decode_action(a)
                                cand_next = (state[0] + dc, state[1] + dp, state[2] + ds)
                                if is_valid_state(cand_next):
                                    legal_actions.append(a)

                            if not legal_actions:
                                break

                            if random.random() < agent_isolated.epsilon:
                                action = random.choice(legal_actions)
                            else:
                                norm_state = normalize_state(state)
                                state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
                                with torch.no_grad():
                                    q_values = agent_isolated.model(state_tensor)[0]
                                action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                            dc, dp, ds = agent_isolated.decode_action(action)
                            next_state = clip_state((state[0] + dc, state[1] + dp, state[2] + ds))

                            triggered = execute_Tr(*next_state)
                            reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                            if target_path.issubset(triggered):
                                reward += 2.0

                            done = (step == N_STEPS - 1)

                            td_error = agent_isolated.store_transition(state, action, reward, next_state, done)

                            priority_stats = agent_isolated.replay_buffer.get_priority_statistics()
                            current_priority = priority_stats['mean_priority'] if priority_stats else 0

                            prioritized_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                    current_priority, None)

                            episode_similarities.append(jaccard_similarity(triggered, target_path))
                            episode_td_errors.append(td_error)

                            if prev_reward is not None:
                                prioritized_metrics.record_action_improvement(reward, prev_reward)

                            prev_state = state
                            prev_triggered = triggered
                            prev_reward = reward
                            state = next_state
                            episode_reward += reward

                    if len(agent_isolated.replay_buffer) >= batch_size:
                        agent_isolated.train(batch_size)
                        global_replay_count += 1
                        if global_replay_count % 2 == 0:
                            agent_isolated.update_target_model()

                # Stage1
                if stage1_samples:
                    for batch_idx in range(N_BATCHES_STAGE1):
                        batch_start = batch_idx * BATCH_SIZE
                        batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES_STAGE1)

                        for sample_idx in range(batch_start, batch_end):
                            if sample_idx >= len(stage1_samples):
                                break

                            stage1_state_tuple, _, _, _ = stage1_samples[sample_idx]
                            state = stage1_state_tuple
                            prev_state = None
                            prev_triggered = None
                            prev_reward = None

                            for step in range(N_STEPS):
                                legal_actions = []
                                for a in range(agent_isolated.action_dim):
                                    dc, dp, ds = agent_isolated.decode_action(a)
                                    cand_next = (state[0] + dc, state[1] + dp, state[2] + ds)
                                    if is_valid_state(cand_next):
                                        legal_actions.append(a)

                                if not legal_actions:
                                    break

                                if random.random() < agent_isolated.epsilon:
                                    action = random.choice(legal_actions)
                                else:
                                    norm_state = normalize_state(state)
                                    state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
                                    with torch.no_grad():
                                        q_values = agent_isolated.model(state_tensor)[0]
                                    action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                                dc, dp, ds = agent_isolated.decode_action(action)
                                next_state = clip_state((state[0] + dc, state[1] + dp, state[2] + ds))

                                triggered = execute_Tr(*next_state)
                                reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                                reward *= 0.8

                                done = (step == N_STEPS - 1)

                                td_error = agent_isolated.store_transition(state, action, reward, next_state, done)

                                priority_stats = agent_isolated.replay_buffer.get_priority_statistics()
                                current_priority = priority_stats['mean_priority'] if priority_stats else 0

                                prioritized_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                        current_priority, None)

                                episode_similarities.append(jaccard_similarity(triggered, target_path))
                                episode_td_errors.append(td_error)

                                if prev_reward is not None:
                                    prioritized_metrics.record_action_improvement(reward, prev_reward)

                                prev_state = state
                                prev_triggered = triggered
                                prev_reward = reward
                                state = next_state
                                episode_reward += reward

                        if len(agent_isolated.replay_buffer) >= batch_size:
                            agent_isolated.train(batch_size)
                            global_replay_count += 1
                            if global_replay_count % 2 == 0:
                                agent_isolated.update_target_model()

            trained_paths.add(path_idx)

        avg_similarity = np.mean(episode_similarities) if episode_similarities else 0
        avg_td_error = np.mean(episode_td_errors) if episode_td_errors else 0
        priority_stats = agent_isolated.replay_buffer.get_priority_statistics()

        prioritized_metrics.record_episode_metrics(episode, episode_reward, avg_similarity, avg_td_error,
                                                   agent_isolated.epsilon, "isolated", priority_stats)

        if len(trained_paths) == len(isolated_group):
            break

    return agent_isolated

# ==================== Excel 导出（性能指标）====================
def append_performance_metrics_to_excel(metrics_collector, filepath, run_number):
    avg_td_error = np.mean(metrics_collector.td_errors) if metrics_collector.td_errors else 0
    total_reward = metrics_collector.total_reward
    action_improvement_rate = np.mean(metrics_collector.action_improvements) if metrics_collector.action_improvements else 0
    avg_final_similarity = np.mean(metrics_collector.final_output_similarities) if metrics_collector.final_output_similarities else 0
    avg_episode_reward = np.mean(metrics_collector.episode_rewards) if metrics_collector.episode_rewards else 0
    reward_std = np.std(metrics_collector.episode_rewards) if metrics_collector.episode_rewards else 0
    training_time = metrics_collector.end_time - metrics_collector.start_time if metrics_collector.end_time and metrics_collector.start_time else 0
    avg_memory_usage = np.mean(metrics_collector.episode_memory_usage) if metrics_collector.episode_memory_usage else 0
    per_step_time = training_time / metrics_collector.step_count * 1000 if metrics_collector.step_count > 0 else 0
    avg_priority = np.mean(metrics_collector.priority_statistics) if metrics_collector.priority_statistics else 0
    avg_importance_weight = np.mean(metrics_collector.importance_weights) if metrics_collector.importance_weights else 0

    new_row = {
        'Run': f"Run {run_number}",
        'Final Avg Similarity': f"{avg_final_similarity:.4f}",
        'Total Reward': f"{total_reward:,.2f}",
        'Avg Episode Reward': f"{avg_episode_reward:,.4f}",
        'Reward Std': f"{reward_std:,.4f}",
        'TD Error': f"{avg_td_error:.4f}",
        'Action Improvement Rate': f"{action_improvement_rate:.4f}",
        'Improvement %': f"{action_improvement_rate * 100:.2f}%",
        'Total Steps': f"{metrics_collector.step_count:,}",
        'Final Samples': f"{len(metrics_collector.final_output_similarities)}",
        'Training Time (s)': f"{training_time:.2f}",
        'Training Time (min)': f"{training_time / 60:.2f}",
        'Avg Memory (MB)': f"{avg_memory_usage:.2f}",
        'Per Step (ms)': f"{per_step_time:.2f}",
        'Avg Priority': f"{avg_priority:.4f}",
        'Avg IS Weight': f"{avg_importance_weight:.4f}"
    }

    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    if os.path.exists(filepath):
        df = pd.read_excel(filepath, sheet_name='Metric')
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    else:
        df = pd.DataFrame([new_row])

    with pd.ExcelWriter(filepath, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, sheet_name='Metric', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Metric']
        worksheet.column_dimensions['A'].width = 15
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
            worksheet.column_dimensions[col].width = 18
        header_font = Font(bold=True, size=11)
        for cell in worksheet[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    print(f"Run {run_number} performance metrics saved to: {filepath}")

# ==================== Excel 导出（最终样本）====================
def append_final_samples_to_excel(agent_similar, agent_isolated, similar_group, isolated_group, targetPaths, filepath,
                                  run_number):
    new_samples = []

    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_similar.replay_buffer.get_high_reward_samples(target_path, num_samples=20)

        for state_tuple, reward, sim, triggered in high_reward_samples:
            new_samples.append({
                'Run': f"Run {run_number}",
                'Path Group': 'Similar path group',
                'Path ID': path_idx + 1,
                'Config_Depth': state_tuple[0],
                'Param_Count': state_tuple[1],
                'Security_Weight': state_tuple[2],
                'Similarity': f"{sim:.4f}",
                'Reward': f"{reward:.2f}",
                'Triggered Count': len(triggered),
                'Target Count': len(target_path),
                'Triggered Rules': str(sorted(triggered)),
                'Target Rules': str(sorted(target_path))
            })

    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)

        for state_tuple, reward, sim, triggered in high_reward_samples:
            new_samples.append({
                'Run': f"Run {run_number}",
                'Path Group': 'Isolated path group',
                'Path ID': path_idx + 1,
                'Config_Depth': state_tuple[0],
                'Param_Count': state_tuple[1],
                'Security_Weight': state_tuple[2],
                'Similarity': f"{sim:.4f}",
                'Reward': f"{reward:.2f}",
                'Triggered Count': len(triggered),
                'Target Count': len(target_path),
                'Triggered Rules': str(sorted(triggered)),
                'Target Rules': str(sorted(target_path))
            })

    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    if os.path.exists(filepath):
        df = pd.read_excel(filepath, sheet_name='Final Samples')
        df = pd.concat([df, pd.DataFrame(new_samples)], ignore_index=True)
    else:
        df = pd.DataFrame(new_samples)

    with pd.ExcelWriter(filepath, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, sheet_name='Final Samples', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Final Samples']

        column_widths = {
            'A': 12, 'B': 15, 'C': 12, 'D': 14, 'E': 14, 'F': 16,
            'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 40, 'L': 40
        }
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        header_font = Font(bold=True, size=11)
        for cell in worksheet[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    print(f"Run {run_number} final samples saved to: {filepath}")

# ==================== 单次实验主流程 ====================
def run_single_experiment(run_number, results_save_dir):
    print(f"\n{'=' * 80}")
    print(f"Start run  {run_number}  (Prioritized DQN - Validation Rules)")
    print(f"{'=' * 80}\n")

    prioritized_metrics.reset()
    prioritized_metrics.start_training()

    model_path_similar = os.path.join(results_save_dir, f"run{run_number}_similar.pth")
    model_path_isolated = os.path.join(results_save_dir, f"run{run_number}_isolated.pth")
    path_documents = r"D:\Experiment\CNN\DQNNEW\path_samples_prioritized"

    similar_group, isolated_group = group_paths_by_similarity(targetPaths)

    # 第一次运行生成相似组样本
    if run_number == 1:
        generate_samples_for_similar_paths(similar_group, num_total=2000, top_k=200)

    replay_buffer = PrioritizedExperienceReplay(capacity=10000, alpha=0.6, beta_start=0.4, beta_frames=100000)
    state_dim = 3
    action_dim = 30
    agent = PrioritizedDQNAgent(state_dim, action_dim, replay_buffer)

    agent = prioritized_generate_and_train_for_similar_paths(agent, similar_group, path_documents, episodes=500,
                                                             batch_size=32, is_isolated=False)

    os.makedirs(os.path.dirname(model_path_similar), exist_ok=True)
    torch.save({
        'model_state_dict': agent.model.state_dict(),
        'optimizer_state_dict': agent.optimizer.state_dict(),
        'epsilon': agent.epsilon
    }, model_path_similar)

    # 第一次运行生成孤立组样本
    if run_number == 1:
        generate_samples_for_isolated_paths_prioritized(agent, isolated_group, num_total=2000, top_k=200)

    isolated_replay_buffer = PrioritizedExperienceReplay(capacity=15000, alpha=0.6, beta_start=0.4, beta_frames=100000)
    agent_isolated = PrioritizedDQNAgent(state_dim, action_dim, isolated_replay_buffer)

    try:
        checkpoint = torch.load(model_path_similar)
        agent_isolated.model.load_state_dict(checkpoint['model_state_dict'])
        agent_isolated.target_model.load_state_dict(checkpoint['model_state_dict'])
        agent_isolated.epsilon = checkpoint.get('epsilon', 0.5)
    except Exception as e:
        pass

    agent_isolated = prioritized_generate_and_train_for_isolated_paths(
        agent_similar=agent,
        agent_isolated=agent_isolated,
        similar_group=similar_group,
        isolated_group=isolated_group,
        path_documents=path_documents,
        episodes=500,
        batch_size=32,
        is_isolated=True
    )

    os.makedirs(os.path.dirname(model_path_isolated), exist_ok=True)
    torch.save({
        'model_state_dict': agent_isolated.model.state_dict(),
        'optimizer_state_dict': agent_isolated.optimizer.state_dict(),
        'epsilon': agent_isolated.epsilon
    }, model_path_isolated)

    prioritized_metrics.end_training()

    # 记录最终样本相似度
    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, _, sim, triggered in high_reward_samples:
            prioritized_metrics.record_final_output_sample(triggered, target_path)

    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, _, sim, triggered in high_reward_samples:
            prioritized_metrics.record_final_output_sample(triggered, target_path)

    # 保存 Excel
    performance_excel_path = os.path.join(results_save_dir, "Metrics_Summary.xlsx")
    samples_excel_path = os.path.join(results_save_dir, "Final_Samples.xlsx")

    append_performance_metrics_to_excel(prioritized_metrics, performance_excel_path, run_number)
    append_final_samples_to_excel(agent, agent_isolated, similar_group, isolated_group, targetPaths, samples_excel_path,
                                  run_number)

    avg_similarity = np.mean(prioritized_metrics.final_output_similarities) if prioritized_metrics.final_output_similarities else 0
    training_time = prioritized_metrics.end_time - prioritized_metrics.start_time
    avg_priority = np.mean(prioritized_metrics.priority_statistics) if prioritized_metrics.priority_statistics else 0
    print(f"\nRun  {run_number} completed:")
    print(f"  Average Similarity: {avg_similarity:.4f}")
    print(f"  Training Time: {training_time:.2f} seconds")
    print(f"  Total Steps: {prioritized_metrics.step_count}")
    print(f"  Avg Priority: {avg_priority:.4f}")

# ==================== 主程序 ====================
if __name__ == "__main__":
    results_save_dir = r"D:\Experiment\CNN\DQNNEW\results\prioritized_validation_results"
    os.makedirs(results_save_dir, exist_ok=True)

    NUM_RUNS = 20

    print("=" * 80)
    print(f"Running {NUM_RUNS} runs of Prioritized DQN (Validation Rules)")
    print(f"State ranges: config_depth[{CONFIG_DEPTH_MIN},{CONFIG_DEPTH_MAX}], "
          f"param_count[{PARAM_COUNT_MIN},{PARAM_COUNT_MAX}], "
          f"security_weight[{SEC_WEIGHT_MIN},{SEC_WEIGHT_MAX}]")
    print("=" * 80)

    for run in range(1, NUM_RUNS + 1):
        try:
            run_single_experiment(run, results_save_dir)
        except Exception as e:
            print(f"\nRun  {run}  failed: {str(e)}")
            import traceback
            traceback.print_exc()
            continue

    print("\n" + "=" * 80)
    print(f"All {NUM_RUNS} runs completed.")
    print(f"Results saved to: {results_save_dir}")
    print("=" * 80)