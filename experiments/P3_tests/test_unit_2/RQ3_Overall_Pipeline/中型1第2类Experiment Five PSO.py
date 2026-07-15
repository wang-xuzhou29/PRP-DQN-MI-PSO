import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """安全除法"""
    if denominator == 0:
        return default
    return numerator / denominator


# === 执行验证规则函数（第二个单元测试） ===
def execute_validation_rules(a):
    """
    参数 a: (config_depth, param_count, security_weight)
    返回: 触发的规则编号集合
    """
    config_depth, param_count, security_weight = float(a[0]), int(a[1]), float(a[2])
    triggered = set()
    b = {}

    # 原始规则逻辑（完整保留）
    if (config_depth >= 5) != (config_depth >= 5.8):
        b[0] = 1
        triggered.add(1)
    if (config_depth >= 5) != (config_depth == 5):
        b[1] = 2
        triggered.add(2)

    if config_depth >= 5:
        if (param_count >= 100 and security_weight >= 0.8) != (param_count != 100 and security_weight >= 0.8):
            b[2] = 3
            triggered.add(3)
        if (param_count >= 100 and security_weight >= 0.8) != (param_count == 100 and security_weight >= 0.8):
            b[3] = 4
            triggered.add(4)
        if (param_count >= 100 and security_weight >= 0.8) != (param_count >= 100 and security_weight != 0.8):
            b[4] = 5
            triggered.add(5)
        if (param_count >= 100 and security_weight >= 0.8) != (param_count >= 100 and security_weight == 0.8):
            b[5] = 6
            triggered.add(6)
        if (param_count >= 100 and security_weight >= 0.8) != (param_count >= 100 or security_weight >= 0.8):
            b[6] = 7
            triggered.add(7)

        if param_count >= 100 and security_weight >= 0.8:
            if (security_weight >= 0.95) != (security_weight == 0.95):
                b[7] = 8
                triggered.add(8)
            if (security_weight >= 0.95) != (security_weight != 0.95):
                b[8] = 9
                triggered.add(9)
            if (security_weight >= 0.9) != (security_weight != 0.9):
                b[9] = 10
                triggered.add(10)
            if (security_weight >= 0.9) != (security_weight == 0.9):
                b[10] = 11
                triggered.add(11)

        if (param_count >= 50 and security_weight >= 0.6) != (param_count != 50 and security_weight >= 0.6):
            b[11] = 12
            triggered.add(12)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count == 50 and security_weight >= 0.6):
            b[12] = 13
            triggered.add(13)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 60 and security_weight >= 0.6):
            b[13] = 14
            triggered.add(14)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 50 and security_weight != 0.6):
            b[14] = 15
            triggered.add(15)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 50 and security_weight == 0.6):
            b[15] = 16
            triggered.add(16)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 53 and security_weight >= 0.6):
            b[16] = 17
            triggered.add(17)

        elif param_count >= 50 and security_weight >= 0.6:
            if (security_weight >= 0.8) != (security_weight != 0.8):
                b[17] = 18
                triggered.add(18)
            if (security_weight >= 0.8) != (security_weight == 0.8):
                b[18] = 19
                triggered.add(19)

        if (param_count >= 25) != (param_count != 25):
            b[19] = 20
            triggered.add(20)
        if (param_count >= 25) != (param_count >= 35):
            b[20] = 21
            triggered.add(21)

    if (config_depth >= 3) != (config_depth != 3):
        b[21] = 22
        triggered.add(22)
    if (config_depth >= 3) != (config_depth >= 3.9):
        b[22] = 23
        triggered.add(23)

    elif config_depth >= 3:
        if (param_count >= 75 and security_weight >= 0.7) != (param_count >= 75 or security_weight >= 0.7):
            b[23] = 24
            triggered.add(24)
        if (param_count >= 75 and security_weight >= 0.7) != (param_count == 75 and security_weight >= 0.7):
            b[24] = 25
            triggered.add(25)
        if (param_count >= 75 and security_weight >= 0.7) != (param_count >= 75 and security_weight != 0.7):
            b[25] = 26
            triggered.add(26)
        if (param_count >= 75 and security_weight >= 0.7) != (param_count >= 75 and security_weight == 0.7):
            b[26] = 27
            triggered.add(27)
        if (param_count >= 75 and security_weight >= 0.7) != (param_count != 75 and security_weight >= 0.7):
            b[27] = 28
            triggered.add(28)

        if param_count >= 75 and security_weight >= 0.7:
            if (security_weight >= 0.9) != (security_weight != 0.9):
                b[28] = 29
                triggered.add(29)
            if (security_weight >= 0.9) != (security_weight == 0.9):
                b[29] = 30
                triggered.add(30)
            if (security_weight >= 0.8) != (security_weight != 0.8):
                b[30] = 31
                triggered.add(31)
            if (security_weight >= 0.8) != (security_weight == 0.8):
                b[31] = 32
                triggered.add(32)

        if (param_count >= 40 and security_weight >= 0.5) != (param_count >= 40 or security_weight >= 0.5):
            b[32] = 33
            triggered.add(33)
        if (param_count >= 40 and security_weight >= 0.5) != (param_count >= 40 and security_weight != 0.5):
            b[33] = 34
            triggered.add(34)
        if (param_count >= 40 and security_weight >= 0.5) != (param_count >= 40 and security_weight == 0.5):
            b[34] = 35
            triggered.add(35)
        if (param_count >= 40 and security_weight >= 0.5) != (param_count != 40 and security_weight >= 0.5):
            b[35] = 36
            triggered.add(36)
        if (param_count >= 40 and security_weight >= 0.5) != (param_count == 40 and security_weight >= 0.5):
            b[36] = 37
            triggered.add(37)

    if (config_depth >= 2) != (config_depth != 2):
        b[37] = 38
        triggered.add(38)
    if (config_depth >= 2) != (config_depth >= 3):
        b[38] = 39
        triggered.add(39)

    elif config_depth >= 2:
        if (param_count >= 50 and security_weight >= 0.6) != (param_count != 50 and security_weight >= 0.6):
            b[39] = 40
            triggered.add(40)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count == 50 and security_weight >= 0.6):
            b[40] = 41
            triggered.add(41)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 50 or security_weight >= 0.6):
            b[41] = 42
            triggered.add(42)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 50 and security_weight == 0.6):
            b[42] = 43
            triggered.add(43)
        if (param_count >= 50 and security_weight >= 0.6) != (param_count >= 50 and security_weight != 0.6):
            b[43] = 44
            triggered.add(44)

        if (param_count >= 30) != (param_count != 30):
            b[44] = 45
            triggered.add(45)
        if (param_count >= 30) != (param_count >= 45):
            b[45] = 46
            triggered.add(46)

    # 参数数量维度分析
    if (param_count >= 150) != (param_count != 150):
        b[46] = 47
        triggered.add(47)
    if (param_count >= 150) != (param_count >= 100):
        b[47] = 48
        triggered.add(48)

    if param_count >= 150:
        if (config_depth >= 4 and security_weight >= 0.8) != (config_depth >= 4 or security_weight >= 0.8):
            b[48] = 49
            triggered.add(49)
        if (config_depth >= 4 and security_weight >= 0.8) != (config_depth == 4 and security_weight >= 0.8):
            b[49] = 50
            triggered.add(50)
        if (config_depth >= 4 and security_weight >= 0.8) != (config_depth != 4 and security_weight >= 0.8):
            b[50] = 51
            triggered.add(51)
        if (config_depth >= 4 and security_weight >= 0.8) != (config_depth >= 4 and security_weight == 0.8):
            b[51] = 52
            triggered.add(52)
        if (config_depth >= 4 and security_weight >= 0.8) != (config_depth >= 4 and security_weight != 0.8):
            b[52] = 53
            triggered.add(53)

        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth >= 3 and security_weight != 0.7):
            b[53] = 54
            triggered.add(54)
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth >= 3 and security_weight == 0.7):
            b[54] = 55
            triggered.add(55)
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth >= 3 or security_weight >= 0.7):
            b[55] = 56
            triggered.add(56)
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth != 3 and security_weight >= 0.7):
            b[56] = 57
            triggered.add(57)
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth == 3 and security_weight >= 0.7):
            b[57] = 58
            triggered.add(58)

    if (param_count >= 100) != (param_count >= 125):
        b[58] = 59
        triggered.add(59)
    if (param_count >= 100) != (param_count >= 200):
        b[59] = 60
        triggered.add(60)

    elif param_count >= 100:
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth >= 5 and security_weight >= 0.81):
            b[60] = 61
            triggered.add(61)
        if (config_depth >= 3 and security_weight >= 0.7) != (config_depth >= 5 and security_weight >= 0.7):
            b[61] = 62
            triggered.add(62)
        if (security_weight >= 0.6) != (security_weight == 0.6):
            b[62] = 63
            triggered.add(63)
        if (security_weight >= 0.6) != (security_weight != 0.6):
            b[63] = 64
            triggered.add(64)

    if (param_count >= 50) != (param_count != 50):
        b[64] = 65
        triggered.add(65)
    if (param_count >= 50) != (param_count >= 74):
        b[65] = 66
        triggered.add(66)

    elif param_count >= 50:
        if (security_weight >= 0.8) != (security_weight != 0.8):
            b[66] = 67
            triggered.add(67)
        if (security_weight >= 0.8) != (security_weight == 0.8):
            b[67] = 68
            triggered.add(68)
        if (security_weight >= 0.6) != (security_weight != 0.6):
            b[68] = 69
            triggered.add(69)
        if (security_weight >= 0.6) != (security_weight == 0.6):
            b[69] = 70
            triggered.add(70)

    # 安全权重维度检查
    if (security_weight >= 0.9) != (security_weight != 0.9):
        b[70] = 71
        triggered.add(71)
    if (security_weight >= 0.9) != (security_weight == 0.9):
        b[71] = 72
        triggered.add(72)

    if security_weight >= 0.9:
        if (config_depth >= 4 and param_count >= 80) != (config_depth != 4 and param_count >= 80):
            b[72] = 73
            triggered.add(73)
        if (config_depth >= 4 and param_count >= 80) != (config_depth == 4 and param_count >= 80):
            b[73] = 74
            triggered.add(74)
        if (config_depth >= 4 and param_count >= 80) != (config_depth >= 4 and param_count != 80):
            b[74] = 75
            triggered.add(75)
        if (config_depth >= 4 and param_count >= 80) != (config_depth >= 4 and param_count == 80):
            b[75] = 76
            triggered.add(76)
        if (config_depth >= 4 and param_count >= 80) != (config_depth >= 4 and param_count >= 88):
            b[76] = 77
            triggered.add(77)
        if (config_depth >= 4 and param_count >= 80) != (config_depth >= 4.4 and param_count >= 80):
            b[77] = 78
            triggered.add(78)

        if (config_depth >= 3 and param_count >= 50) != (config_depth >= 3 and param_count != 50):
            b[78] = 79
            triggered.add(79)
        if (config_depth >= 3 and param_count >= 50) != (config_depth >= 3 and param_count == 50):
            b[79] = 80
            triggered.add(80)
        if (config_depth >= 3 and param_count >= 50) != (config_depth != 3 and param_count >= 50):
            b[80] = 81
            triggered.add(81)
        if (config_depth >= 3 and param_count >= 50) != (config_depth == 3 and param_count >= 50):
            b[81] = 82
            triggered.add(82)
        if (config_depth >= 3 and param_count >= 50) != (config_depth >= 3.6 and param_count >= 50):
            b[82] = 83
            triggered.add(83)
        if (config_depth >= 3 and param_count >= 50) != (config_depth >= 3 and param_count >= 66):
            b[83] = 84
            triggered.add(84)

    if (security_weight >= 0.7) != (security_weight != 0.7):
        b[84] = 85
        triggered.add(85)
    if (security_weight >= 0.7) != (security_weight == 0.7):
        b[85] = 86
        triggered.add(86)

    elif security_weight >= 0.7:
        if (config_depth >= 3 and param_count >= 60) != (config_depth >= 3 and param_count >= 66):
            b[86] = 87
            triggered.add(87)
        if (config_depth >= 3 and param_count >= 60) != (config_depth >= 3 and param_count == 60):
            b[87] = 88
            triggered.add(88)
        if (config_depth >= 3 and param_count >= 60) != (config_depth != 3 and param_count >= 60):
            b[88] = 89
            triggered.add(89)
        if (config_depth >= 3 and param_count >= 60) != (config_depth == 3 and param_count >= 60):
            b[89] = 90
            triggered.add(90)
        if (config_depth >= 3 and param_count >= 60) != (config_depth >= 3 and param_count >= 72):
            b[90] = 91
            triggered.add(91)
        if (config_depth >= 3 and param_count >= 60) != (config_depth >= 3.3 and param_count >= 60):
            b[91] = 92
            triggered.add(92)

        if (param_count >= 40) != (param_count != 40):
            b[92] = 93
            triggered.add(93)
        if (param_count >= 40) != (param_count >= 25):
            b[93] = 94
            triggered.add(94)

    if (security_weight >= 0.5) != (security_weight != 0.5):
        b[94] = 95
        triggered.add(95)
    if (security_weight >= 0.5) != (security_weight == 0.5):
        b[95] = 96
        triggered.add(96)

    elif security_weight >= 0.5:
        if (param_count >= 30) != (param_count >= 80):
            b[96] = 97
            triggered.add(97)
        if (param_count >= 30) != (param_count >= 100):
            b[97] = 98
            triggered.add(98)

    # Missing parameters detection based on dimensions
    if (config_depth < 3) != (config_depth < 4.7):
        b[98] = 99
        triggered.add(99)
    if (config_depth < 3) != (config_depth < 5):
        b[99] = 100
        triggered.add(100)
    if (param_count < 50) != (param_count < 76):
        b[100] = 101
        triggered.add(101)
    if (param_count < 50) != (param_count < 67):
        b[101] = 102
        triggered.add(102)
    if (security_weight < 0.6) != (security_weight == 0.63):
        b[102] = 103
        triggered.add(103)
    if (security_weight < 0.6) != (security_weight != 0.6):
        b[103] = 104
        triggered.add(104)

    return triggered


# 将执行函数统一命名为 execute_Tr（供 PSO 调用）
execute_Tr = execute_validation_rules


# === 适应度函数 ===
def calculate_fitness(position: List[float], target_path: Set[int]) -> float:
    """计算 Jaccard 相似度（若完全包含目标路径则返回 1.0）"""
    generated_path = execute_Tr(position)   # position 为 [config_depth, param_count, security_weight]

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


# === 标准 PSO 类 ===
class BasicPSO:
    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # 修改为第二个单元测试的边界
        self.bounds = bounds if bounds else [(1, 7), (1, 180), (0, 1)]  # config_depth, param_count, security_weight
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        particles = []
        velocities = []
        for _ in range(self.n_particles):
            particle = []
            for i in range(self.dim):
                if i == 0:  # config_depth 整数
                    particle.append(random.randint(self.bounds[i][0], self.bounds[i][1]))
                elif i == 1:  # param_count 整数
                    particle.append(random.randint(self.bounds[i][0], self.bounds[i][1]))
                else:  # security_weight 浮点数，在 [0,1] 内
                    particle.append(random.uniform(self.bounds[i][0], self.bounds[i][1]))
            particles.append(particle)
            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)
        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        new_velocity = []
        new_particle = []
        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))
            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            # 对整数维度取整，浮点维度保留浮点
            if i in [0, 1]:  # config_depth 和 param_count 为整数
                p = round(p)
            # 对浮点维度不取整
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        start_time = time.time()
        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    return {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_Tr(particles[i]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()
                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

        return {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_Tr(gbest_particle),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }


# === 运行多次实验 ===
def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs} 次运行")
    print(f"{'=' * 70}")
    print(f"参数: {n_particles} 个粒子, {max_iterations} 次迭代, {len(target_paths)} 条目标路径")
    print(f"搜索空间: config_depth [1,7], param_count [1,180], security_weight [0,1]")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"--- 第 {run_idx}/{num_runs} 次运行 ---")
        results = {}
        pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

        for i, target_path in enumerate(target_paths):
            print(f"  Path {i+1}: ", end='')
            result = pso.optimize(target_path)
            results[i] = result
            status = "✓" if result['success'] else f"({result['best_fitness']:.3f})"
            print(f"{status} | 耗时 {result['time']:.2f}s | 迭代 {result['iterations']}")

        all_results.append(results)
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f"  本轮成功: {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start
    print(f"{'=' * 70}")
    print(f"全部 {num_runs} 次运行完成 | 总耗时 {total_time:.2f}s")
    print(f"{'=' * 70}\n")
    return all_results


# === 导出 Excel ===
def export_to_excel(all_results, target_paths, filename=None):
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Validation_Results_{timestamp}.xlsx"

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 工作表1：运行统计
    ws1 = wb.active
    ws1.title = "运行统计"
    ws1.sheet_view.showGridLines = False
    headers = ["运行", "成功率", "成功数", "平均适应度", "平均迭代次数", "PSO时间(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = sum(results[i]['time'] for i in range(len(target_paths)))

        row_data = [
            f"Run {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx+1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align
            if run_idx % 2 == 0:
                cell.fill = alternate_fill
            if col == 2 and success_rate == 100.0:
                cell.fill = success_fill
            elif col == 2 and success_rate < 50.0:
                cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results)+1}"

    # 工作表2：路径统计
    ws2 = wb.create_sheet(title="路径统计")
    ws2.sheet_view.showGridLines = False
    headers2 = ["Path ID", "成功数", "成功率", "平均适应度", "平均迭代", "最小迭代", "最大迭代"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]
    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])
        iters = [r[path_idx]['iterations'] for r in all_results]
        row_data = [
            f"Path {path_idx+1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{np.mean(iters):.1f}",
            f"{np.min(iters)}",
            f"{np.max(iters)}"
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx+2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align
            if (path_idx+1) % 2 == 0:
                cell.fill = alternate_fill
            if col == 3 and success_rate == 100.0:
                cell.fill = success_fill
            elif col == 3 and success_rate < 50.0:
                cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths)+1}"

    # 工作表3：详细结果
    ws3 = wb.create_sheet(title="详细结果")
    ws3.sheet_view.showGridLines = False
    headers3 = ["Path", "Run", "(config_depth, param_count, security_weight)", "适应度", "迭代次数", "触发路径"]
    col_widths3 = [10, 10, 30, 12, 12, 55]
    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            result = results[path_idx]
            pos = result['best_particle']
            fitness = result['best_fitness']
            path_set = result['best_path']
            iters = result['iterations']
            row_data = [
                f"Path {path_idx+1}",
                f"Run {run_idx}",
                f"({pos[0]:.3f}, {pos[1]}, {pos[2]:.3f})",
                f"{fitness:.4f}",
                iters,
                str(sorted(list(path_set)))
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = left_align if col == 6 else center_align
                if fitness == 1.0:
                    cell.fill = success_fill
                elif fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill
            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx-1}"

    # 工作表4：目标路径
    ws4 = wb.create_sheet(title="目标路径")
    ws4.sheet_view.showGridLines = False
    headers4 = ["Path ID", "目标路径", "长度"]
    col_widths4 = [12, 60, 12]
    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        row_data = [
            f"Path {path_idx+1}",
            str(sorted(list(target_path))),
            len(target_path)
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx+2, column=col, value=value)
            cell.border = border
            cell.alignment = left_align if col == 2 else center_align
            if (path_idx+1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)
    print(f"\n{'=' * 70}")
    print(f"Excel 已保存: {filename}")
    print(f"{'=' * 70}")
    return filename


# === 主函数 ===
def main():
    # === 目标路径组（第二个单元测试） ===
    targetPaths = [
        # A1
        {3, 7, 8, 11, 12, 14, 16, 17, 19, 22, 23, 24, 28, 30, 32, 35, 37, 40, 43, 47, 49, 51, 55, 57, 61, 62, 63, 65,
         66, 68, 70, 72, 79, 81, 83, 84, 86, 96, 97, 98, 99, 100, 101, 102, 104},
        # A2
        {3, 7, 8, 11, 12, 14, 16, 17, 19, 24, 28, 30, 32, 35, 37, 40, 43, 47, 51, 52, 55, 58, 61, 62, 63, 65, 66, 68,
         70, 72, 75, 79, 82, 84, 86, 96, 97, 98, 99, 100, 101, 102, 104},
        # A3
        {3, 6, 8, 11, 13, 16, 19, 22, 23, 25, 27, 30, 32, 35, 37, 41, 43, 47, 48, 49, 51, 55, 57, 59, 60, 61, 62, 63,
         68, 70, 72, 73, 80, 81, 83, 86, 88, 89, 92, 96, 99, 100, 104},
        # A4
        {3, 7, 8, 11, 12, 19, 21, 24, 28, 30, 32, 33, 36, 40, 42, 45, 46, 47, 51, 52, 55, 58, 61, 62, 63, 65, 68, 70,
         72, 75, 79, 86, 93, 94, 96, 97, 98, 99, 100, 104},
        # A5
        {1, 3, 7, 8, 11, 13, 16, 19, 24, 28, 30, 32, 35, 37, 41, 43, 47, 50, 52, 55, 58, 63, 66, 68, 70, 72, 75, 80, 82,
         84, 86, 87, 90, 91, 96, 97, 98, 101, 102, 104},
        # A6
        {2, 3, 7, 8, 11, 12, 19, 20, 21, 24, 28, 30, 32, 33, 36, 40, 42, 45, 47, 50, 52, 55, 58, 63, 65, 68, 70, 72, 75,
         79, 86, 93, 94, 96, 104},
        # A7
        {1, 9, 10, 15, 18, 24, 26, 29, 31, 33, 34, 42, 44, 47, 49, 53, 54, 56, 64, 67, 69, 71, 74, 75, 77, 80, 82, 85,
         88, 90, 95, 98, 103},
        # A8
        {5, 7, 9, 10, 15, 18, 22, 24, 26, 29, 31, 33, 34, 38, 42, 44, 47, 60, 64, 67, 69, 71, 73, 81, 85, 89, 95, 103},
        # A9
        {4, 6, 8, 11, 13, 16, 19, 25, 27, 30, 32, 35, 37, 41, 43, 47, 51, 52, 55, 58, 60, 61, 62, 63, 68, 70, 72, 73,
         76, 78, 80, 82, 86, 88, 90, 96, 99, 100, 104},
        # A10
        {3, 7, 8, 11, 13, 16, 19, 22, 27, 28, 30, 32, 35, 37, 38, 39, 41, 43, 47, 49, 51, 56, 57, 63, 68, 70, 72, 81,
         86, 89, 96, 97, 98, 101, 104}
    ]

    print("=" * 70)
    print("baseline PSO - 第二个单元测试（配置验证规则）")
    print("=" * 70)

    all_results = run_multiple_experiments(
        targetPaths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, targetPaths)

    print("Program completed")


if __name__ == "__main__":
    main()