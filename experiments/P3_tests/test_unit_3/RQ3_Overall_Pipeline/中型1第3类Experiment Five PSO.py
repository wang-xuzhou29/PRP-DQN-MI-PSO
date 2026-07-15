import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """安全除法"""
    if denominator == 0:
        return default
    return numerator / denominator


# === 执行编排规则函数（第二个单元测试：编排规则） ===
def execute_orchestration_rules(a):
    """
    参数 a: (path_depth, file_count, access_level)
    返回: 触发的规则编号集合
    """
    path_depth, file_count, access_level = float(a[0]), int(a[1]), float(a[2])
    triggered = set()
    b = {}

    # 原始规则逻辑（完整保留）
    if (path_depth >= 10) != (path_depth >= 12):
        b[0] = 1
        triggered.add(1)
    if (path_depth >= 10) != (path_depth == 10):
        b[1] = 2
        triggered.add(2)
    if (path_depth >= 10) != (path_depth >= 8):
        b[2] = 3
        triggered.add(3)

    if path_depth >= 10:
        if (file_count >= 10000 and access_level >= 3) != (file_count >= 10000 and access_level >= 4.3):
            b[3] = 4
            triggered.add(4)
        if (file_count >= 10000 and access_level >= 3) != (file_count >= 10000 and access_level == 3):
            b[4] = 5
            triggered.add(5)
        if (file_count >= 10000 and access_level >= 3) != (file_count >= 10000 and access_level >= 3.2):
            b[5] = 6
            triggered.add(6)
        if (file_count >= 10000 and access_level >= 3) != (file_count == 10000 and access_level >= 3):
            b[6] = 7
            triggered.add(7)
        if (file_count >= 10000 and access_level >= 3) != (file_count >= 10000 and access_level >= 4):
            b[7] = 8
            triggered.add(8)
        if (file_count >= 10000 and access_level >= 3) != (file_count >= 13400 and access_level >= 3.9):
            b[8] = 9
            triggered.add(9)
        if (file_count >= 10000 and access_level >= 3) != (file_count != 10000 and access_level >= 3):
            b[9] = 10
            triggered.add(10)
        if (file_count >= 10000 and access_level >= 3) != (file_count >= 10000 and access_level >= 5):
            b[10] = 11
            triggered.add(11)

        if file_count >= 10000 and access_level >= 3:
            if (access_level == 3 and file_count >= 50000) != (access_level != 3 and file_count >= 50000):
                b[11] = 12
                triggered.add(12)
            if (access_level == 3 and file_count >= 50000) != (access_level >= 3 and file_count >= 50000):
                b[12] = 13
                triggered.add(13)
            if (access_level == 3 and file_count >= 50000) != (access_level <= 3 and file_count >= 50000):
                b[13] = 14
                triggered.add(14)
            if (access_level == 3 and file_count >= 50000) != (access_level == 3 or file_count >= 50000):
                b[14] = 15
                triggered.add(15)

            if (file_count >= 25000) != (file_count != 25000):
                b[15] = 16
                triggered.add(16)
            if (file_count >= 25000) != (file_count >= 30000):
                b[16] = 17
                triggered.add(17)
            if (file_count >= 25000) != (file_count >= 29000):
                b[17] = 18
                triggered.add(18)

        if (file_count >= 5000 and access_level >= 2) != (file_count >= 5000 and access_level >= 2.5):
            b[18] = 19
            triggered.add(19)
        if (file_count >= 5000 and access_level >= 2) != (file_count >= 5000 or access_level >= 3.5):
            b[19] = 20
            triggered.add(20)
        if (file_count >= 5000 and access_level >= 2) != (file_count >= 5000 and access_level != 2):
            b[20] = 21
            triggered.add(21)
        if (file_count >= 5000 and access_level >= 2) != (file_count != 5000 and access_level >= 2):
            b[21] = 22
            triggered.add(22)
        if (file_count >= 5000 and access_level >= 2) != (file_count >= 5000 and access_level >= 3):
            b[22] = 23
            triggered.add(23)
        if (file_count >= 5000 and access_level >= 2) != (file_count >= 5000 and access_level >= 3.3):
            b[23] = 24
            triggered.add(24)

        if (file_count >= 1000) != (file_count >= 43000):
            b[24] = 25
            triggered.add(25)
        if (file_count >= 1000) != (file_count >= 25000):
            b[25] = 26
            triggered.add(26)
        if (file_count >= 1000) != (file_count >= 50000):
            b[26] = 27
            triggered.add(27)

    if (path_depth >= 6) != (path_depth >= 7):
        b[27] = 28
        triggered.add(28)
    if (path_depth >= 6) != (path_depth != 6):
        b[28] = 29
        triggered.add(29)
    if (path_depth >= 6) != (path_depth >= 7):
        b[29] = 30
        triggered.add(30)

    elif path_depth >= 6:
        if (file_count >= 20000 and access_level >= 3) != (file_count >= 20000 and access_level >= 4):
            b[30] = 31
            triggered.add(31)
        if (file_count >= 20000 and access_level >= 3) != (file_count >= 20000 and access_level != 3):
            b[31] = 32
            triggered.add(32)
        if (file_count >= 20000 and access_level >= 3) != (file_count >= 20000 and access_level == 3):
            b[32] = 33
            triggered.add(33)
        if (file_count >= 20000 and access_level >= 3) != (file_count == 20000 and access_level >= 3):
            b[33] = 34
            triggered.add(34)
        if (file_count >= 20000 and access_level >= 3) != (file_count != 20000 and access_level >= 3):
            b[34] = 35
            triggered.add(35)

        if (file_count >= 8000 and access_level >= 2) != (file_count >= 54000 and access_level >= 2):
            b[35] = 36
            triggered.add(36)
        if (file_count >= 8000 and access_level >= 2) != (file_count != 8000 and access_level >= 2):
            b[36] = 37
            triggered.add(37)
        if (file_count >= 8000 and access_level >= 2) != (file_count >= 45000 and access_level >= 2):
            b[37] = 38
            triggered.add(38)
        if (file_count >= 8000 and access_level >= 2) != (file_count >= 8000 and access_level != 2):
            b[38] = 39
            triggered.add(39)
        if (file_count >= 8000 and access_level >= 2) != (file_count >= 8000 and access_level >= 3.3):
            b[39] = 40
            triggered.add(40)
        if (file_count >= 8000 and access_level >= 2) != (file_count >= 8000 and access_level >= 2.5):
            b[40] = 41
            triggered.add(41)
        if (file_count >= 8000 and access_level >= 2) != (file_count >= 80500 and access_level >= 2):
            b[41] = 42
            triggered.add(42)

        elif file_count >= 8000 and access_level >= 2:
            if (access_level == 3) != (access_level <= 3):
                b[42] = 43
                triggered.add(43)
            if (access_level == 3) != (access_level >= 3):
                b[43] = 44
                triggered.add(44)

        if (file_count >= 2000) != (file_count != 2000):
            b[44] = 45
            triggered.add(45)
        if (file_count >= 2000) != (file_count >= 62000):
            b[45] = 46
            triggered.add(46)

    if (path_depth >= 3) != (path_depth != 3):
        b[46] = 47
        triggered.add(47)
    if (path_depth >= 3) != (path_depth >= 3.5):
        b[47] = 48
        triggered.add(48)

    elif path_depth >= 3:
        if (file_count >= 15000 and access_level >= 2) != (file_count != 15000 and access_level >= 2):
            b[48] = 49
            triggered.add(49)
        if (file_count >= 15000 and access_level >= 2) != (file_count == 15000 and access_level >= 2):
            b[49] = 50
            triggered.add(50)
        if (file_count >= 15000 and access_level >= 2) != (file_count >= 15000 and access_level != 2):
            b[50] = 51
            triggered.add(51)
        if (file_count >= 15000 and access_level >= 2) != (file_count >= 15000 and access_level == 2):
            b[51] = 52
            triggered.add(52)
        if (file_count >= 15000 and access_level >= 2) != (file_count >= 15000 and access_level >= 2.5):
            b[52] = 53
            triggered.add(53)
        if (file_count >= 15000 and access_level >= 2) != (file_count >= 17500 and access_level >= 2):
            b[53] = 54
            triggered.add(54)
        if (file_count >= 15000 and access_level >= 2) != (file_count >= 33330 and access_level >= 2):
            b[54] = 55
            triggered.add(55)

        if file_count >= 15000 and access_level >= 2:
            if (access_level == 3) != (access_level >= 3):
                b[55] = 56
                triggered.add(56)
            if (access_level == 3) != (access_level <= 3):
                b[56] = 57
                triggered.add(57)

        if (file_count >= 5000) != (file_count != 5000):
            b[57] = 58
            triggered.add(58)
        if (file_count >= 5000) != (file_count >= 55000):
            b[58] = 59
            triggered.add(59)

    # 文件数量维度处理
    if (file_count >= 100000) != (file_count >= 90000):
        b[59] = 60
        triggered.add(60)
    if (file_count >= 100000) != (file_count == 100000):
        b[60] = 61
        triggered.add(61)

    if file_count >= 100000:
        if (path_depth >= 8 and access_level >= 3) != (path_depth >= 8 and access_level != 3):
            b[61] = 62
            triggered.add(62)
        if (path_depth >= 8 and access_level >= 3) != (path_depth >= 8 and access_level == 3):
            b[62] = 63
            triggered.add(63)
        if (path_depth >= 8 and access_level >= 3) != (path_depth == 8 and access_level >= 3):
            b[63] = 64
            triggered.add(64)
        if (path_depth >= 8 and access_level >= 3) != (path_depth != 8 and access_level >= 3):
            b[64] = 65
            triggered.add(65)
        if (path_depth >= 8 and access_level >= 3) != (path_depth >= 9 and access_level >= 3):
            b[65] = 66
            triggered.add(66)
        if (path_depth >= 8 and access_level >= 3) != (path_depth >= 11 and access_level >= 3):
            b[66] = 67
            triggered.add(67)
        if (path_depth >= 8 and access_level >= 3) != (path_depth >= 15 and access_level >= 3):
            b[67] = 68
            triggered.add(68)

        if (path_depth >= 5 and access_level >= 2) != (path_depth != 5 and access_level >= 2):
            b[68] = 69
            triggered.add(69)
        if (path_depth >= 5 and access_level >= 2) != (path_depth == 5 and access_level >= 2):
            b[69] = 70
            triggered.add(70)
        if (path_depth >= 5 and access_level >= 2) != (path_depth >= 5 and access_level >= 2.9):
            b[70] = 71
            triggered.add(71)
        if (path_depth >= 5 and access_level >= 2) != (path_depth >= 5 and access_level != 2):
            b[71] = 72
            triggered.add(72)
        if (path_depth >= 5 and access_level >= 2) != (path_depth >= 5 and access_level == 2):
            b[72] = 73
            triggered.add(73)
        if (path_depth >= 5 and access_level >= 2) != (path_depth >= 5 and access_level >= 2.5):
            b[73] = 74
            triggered.add(74)
        if (path_depth >= 5 and access_level >= 2) != (path_depth >= 7 and access_level >= 2):
            b[74] = 75
            triggered.add(75)

    if (file_count >= 50000) != (file_count != 50000):
        b[75] = 76
        triggered.add(76)
    if (file_count >= 50000) != (file_count == 50000):
        b[76] = 77
        triggered.add(77)

    elif file_count >= 50000:
        if (access_level >= 3) != (access_level != 3):
            b[77] = 78
            triggered.add(78)
        if (access_level >= 3) != (access_level == 3):
            b[78] = 79
            triggered.add(79)

    if (file_count >= 10000) != (file_count != 10000):
        b[79] = 80
        triggered.add(80)
    if (file_count >= 10000) != (file_count == 10000):
        b[80] = 81
        triggered.add(81)

    # 访问级别维度的扫描序列规划
    if (access_level == 3) != (access_level <= 3):
        b[81] = 82
        triggered.add(82)
    if (access_level == 3) != (access_level >= 3):
        b[82] = 83
        triggered.add(83)

    if access_level == 3:
        if (path_depth >= 8 and file_count >= 20000) != (path_depth != 8 and file_count >= 20000):
            b[83] = 84
            triggered.add(84)
        if (path_depth >= 8 and file_count >= 20000) != (path_depth == 8 and file_count >= 20000):
            b[84] = 85
            triggered.add(85)
        if (path_depth >= 8 and file_count >= 20000) != (path_depth >= 7 and file_count >= 20000):
            b[85] = 86
            triggered.add(86)
        if (path_depth >= 8 and file_count >= 20000) != (path_depth >= 8 and file_count != 20000):
            b[86] = 87
            triggered.add(87)
        if (path_depth >= 8 and file_count >= 20000) != (path_depth >= 8 and file_count == 20000):
            b[87] = 88
            triggered.add(88)
        if (path_depth >= 8 and file_count >= 20000) != (path_depth >= 11 and file_count == 20000):
            b[88] = 89
            triggered.add(89)
        if (path_depth >= 8 and file_count >= 20000) != (path_depth >= 8 and file_count == 25500):
            b[89] = 90
            triggered.add(90)

        if (path_depth >= 5) != (path_depth != 5):
            b[90] = 91
            triggered.add(91)
        if (path_depth >= 5) != (path_depth >= 3.5):
            b[91] = 92
            triggered.add(92)
        if (path_depth >= 5) != (path_depth >= 6):
            b[92] = 93
            triggered.add(93)

    if (access_level == 2) != (access_level >= 2):
        b[93] = 94
        triggered.add(94)
    if (access_level == 2) != (access_level <= 2):
        b[94] = 95
        triggered.add(95)

    return triggered


# 将执行函数统一命名为 execute_Tr（供 PSO 调用）
execute_Tr = execute_orchestration_rules


# === 适应度函数 ===
def calculate_fitness(position: List[float], target_path: Set[int]) -> float:
    """计算 Jaccard 相似度（若完全包含目标路径则返回 1.0）"""
    generated_path = execute_Tr(position)   # position 为 [path_depth, file_count, access_level]

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


# === 标准 PSO 类 ===
class BasicPSO:
    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # 修改为编排规则的边界
        self.bounds = bounds if bounds else [(1, 15), (1, 110000), (1, 4)]  # path_depth, file_count, access_level
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        particles = []
        velocities = []
        for _ in range(self.n_particles):
            particle = []
            for i in range(self.dim):
                if i == 1:  # file_count 为整数
                    particle.append(random.randint(self.bounds[i][0], self.bounds[i][1]))
                else:  # path_depth 和 access_level 为浮点数
                    particle.append(random.uniform(self.bounds[i][0], self.bounds[i][1]))
            particles.append(particle)
            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)
        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        new_velocity = []
        new_particle = []
        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))
            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            # file_count 为整数，取整；其他维度保留浮点
            if i == 1:
                p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        start_time = time.time()
        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    return {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_Tr(particles[i]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()
                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

        return {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_Tr(gbest_particle),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }


# === 运行多次实验 ===
def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs} 次运行")
    print(f"{'=' * 70}")
    print(f"参数: {n_particles} 个粒子, {max_iterations} 次迭代, {len(target_paths)} 条目标路径")
    print(f"搜索空间: path_depth [1,15], file_count [1,110000], access_level [1,4]")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"--- 第 {run_idx}/{num_runs} 次运行 ---")
        results = {}
        pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

        for i, target_path in enumerate(target_paths):
            print(f"  Path {i+1}: ", end='')
            result = pso.optimize(target_path)
            results[i] = result
            status = "✓" if result['success'] else f"({result['best_fitness']:.3f})"
            print(f"{status} | 耗时 {result['time']:.2f}s | 迭代 {result['iterations']}")

        all_results.append(results)
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f"  本轮成功: {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start
    print(f"{'=' * 70}")
    print(f"全部 {num_runs} 次运行完成 | 总耗时 {total_time:.2f}s")
    print(f"{'=' * 70}\n")
    return all_results


# === 导出 Excel ===
def export_to_excel(all_results, target_paths, filename=None):
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Orchestration_Results_{timestamp}.xlsx"

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 工作表1：运行统计
    ws1 = wb.active
    ws1.title = "运行统计"
    ws1.sheet_view.showGridLines = False
    headers = ["运行", "成功率", "成功数", "平均适应度", "平均迭代次数", "PSO时间(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = sum(results[i]['time'] for i in range(len(target_paths)))

        row_data = [
            f"Run {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx+1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align
            if run_idx % 2 == 0:
                cell.fill = alternate_fill
            if col == 2 and success_rate == 100.0:
                cell.fill = success_fill
            elif col == 2 and success_rate < 50.0:
                cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results)+1}"

    # 工作表2：路径统计
    ws2 = wb.create_sheet(title="路径统计")
    ws2.sheet_view.showGridLines = False
    headers2 = ["Path ID", "成功数", "成功率", "平均适应度", "平均迭代", "最小迭代", "最大迭代"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]
    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])
        iters = [r[path_idx]['iterations'] for r in all_results]
        row_data = [
            f"Path {path_idx+1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{np.mean(iters):.1f}",
            f"{np.min(iters)}",
            f"{np.max(iters)}"
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx+2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align
            if (path_idx+1) % 2 == 0:
                cell.fill = alternate_fill
            if col == 3 and success_rate == 100.0:
                cell.fill = success_fill
            elif col == 3 and success_rate < 50.0:
                cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths)+1}"

    # 工作表3：详细结果
    ws3 = wb.create_sheet(title="详细结果")
    ws3.sheet_view.showGridLines = False
    headers3 = ["Path", "Run", "(path_depth, file_count, access_level)", "适应度", "迭代次数", "触发路径"]
    col_widths3 = [10, 10, 30, 12, 12, 55]
    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            result = results[path_idx]
            pos = result['best_particle']
            fitness = result['best_fitness']
            path_set = result['best_path']
            iters = result['iterations']
            row_data = [
                f"Path {path_idx+1}",
                f"Run {run_idx}",
                f"({pos[0]:.3f}, {int(pos[1])}, {pos[2]:.3f})",
                f"{fitness:.4f}",
                iters,
                str(sorted(list(path_set)))
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = left_align if col == 6 else center_align
                if fitness == 1.0:
                    cell.fill = success_fill
                elif fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill
            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx-1}"

    # 工作表4：目标路径
    ws4 = wb.create_sheet(title="目标路径")
    ws4.sheet_view.showGridLines = False
    headers4 = ["Path ID", "目标路径", "长度"]
    col_widths4 = [12, 60, 12]
    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        row_data = [
            f"Path {path_idx+1}",
            str(sorted(list(target_path))),
            len(target_path)
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx+2, column=col, value=value)
            cell.border = border
            cell.alignment = left_align if col == 2 else center_align
            if (path_idx+1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)
    print(f"\n{'=' * 70}")
    print(f"Excel 已保存: {filename}")
    print(f"{'=' * 70}")
    return filename


# === 主函数 ===
def main():
    # === 目标路径组（编排规则，18条） ===
    targetPaths = [
        # A1
        {1, 4, 6, 7, 8, 9, 11, 15, 17, 24, 25, 27, 31, 32, 34, 36, 38, 40, 42, 46, 50, 52, 55, 59, 62, 64, 67, 68, 70,
         73, 76, 78, 81, 85, 88, 89, 90, 94},
        # A2
        {1, 2, 4, 6, 7, 8, 9, 11, 15, 17, 24, 25, 27, 31, 32, 34, 36, 38, 40, 42, 46, 50, 52, 55, 59, 62, 64, 68, 70,
         73, 76, 78, 81, 85, 88, 89, 90, 94},
        # A3
        {3, 4, 5, 7, 11, 17, 25, 27, 33, 34, 36, 38, 42, 44, 46, 50, 52, 55, 56, 59, 63, 65, 66, 67, 68, 70, 73, 76, 79,
         81, 83, 84, 88, 89, 90, 94},
        # A4
        {17, 19, 21, 23, 24, 25, 27, 28, 29, 30, 32, 36, 38, 39, 40, 41, 42, 43, 46, 50, 51, 53, 55, 57, 59, 70, 71, 72,
         74, 75, 76, 78, 81, 82, 84},
        # A5
        {19, 21, 23, 24, 25, 27, 29, 32, 36, 38, 39, 40, 41, 42, 43, 46, 50, 51, 53, 55, 57, 59, 69, 71, 72, 74, 75, 76,
         78, 81, 82, 84, 91, 93},
        # A6
        {16, 19, 21, 23, 24, 25, 26, 27, 29, 36, 38, 39, 40, 41, 42, 43, 46, 50, 51, 53, 55, 57, 59, 69, 71, 72, 74, 75,
         76, 78, 81, 82, 91, 93},
        # A7
        {3, 4, 6, 7, 8, 9, 11, 15, 16, 24, 25, 26, 27, 35, 36, 38, 40, 42, 46, 50, 52, 54, 55, 59, 62, 64, 67, 68, 70,
         73, 76, 78, 81, 87, 94},
        # A8
        {1, 12, 14, 15, 19, 21, 23, 24, 32, 39, 40, 41, 42, 43, 46, 50, 51, 53, 57, 59, 62, 70, 71, 72, 74, 77, 78, 81,
         82, 85, 88, 89, 90},
        # A9
        {3, 4, 5, 7, 11, 12, 13, 15, 33, 34, 42, 44, 46, 50, 52, 56, 59, 63, 65, 66, 67, 68, 70, 73, 77, 79, 81, 83, 84,
         88, 89, 90, 94},
        # A10
        {17, 19, 21, 23, 24, 25, 27, 32, 36, 38, 39, 40, 41, 42, 43, 46, 50, 51, 53, 55, 57, 59, 70, 71, 72, 74, 76, 78,
         81, 82, 84, 86},
        # A11
        {17, 19, 21, 23, 24, 25, 27, 29, 32, 36, 38, 39, 40, 41, 42, 43, 46, 47, 48, 50, 51, 53, 55, 57, 59, 69, 76, 78,
         81, 82, 84, 91},
        # A12
        {16, 19, 21, 23, 24, 25, 26, 27, 28, 29, 30, 36, 38, 39, 40, 41, 42, 43, 46, 49, 57, 59, 70, 71, 72, 74, 75, 76,
         78, 80, 82},
        # A13
        {16, 19, 21, 23, 24, 25, 26, 27, 29, 36, 38, 39, 40, 41, 42, 43, 46, 50, 51, 53, 55, 57, 59, 69, 76, 78, 81, 82,
         91, 92},
        # A14
        {1, 10, 16, 20, 22, 25, 26, 27, 35, 37, 44, 45, 49, 56, 58, 63, 64, 67, 68, 70, 73, 76, 79, 80, 83, 87, 94},
        # A15
        {1, 2, 17, 20, 21, 25, 27, 32, 39, 43, 46, 51, 57, 59, 62, 72, 76, 78, 81, 82, 85, 88, 89, 90, 95},
        # A16
        {4, 5, 7, 11, 17, 18, 25, 27, 29, 33, 34, 36, 38, 42, 44, 46, 50, 52, 55, 56, 59, 65, 69, 73, 75, 76, 79, 81,
         83, 84, 91, 93, 94},
        # A17
        {4, 6, 7, 8, 9, 11, 12, 24, 31, 32, 34, 40, 50, 52, 61, 65, 70, 73, 77, 78, 81, 84, 86, 94},
        # A18
        {1, 2, 4, 6, 7, 8, 9, 11, 12, 24, 31, 32, 34, 40, 50, 52, 60, 62, 64, 68, 70, 73, 77, 78, 81, 85, 88, 89, 90,
         94}
    ]

    print("=" * 70)
    print("baseline PSO - 编排规则单元测试（第二个单元测试）")
    print("=" * 70)

    all_results = run_multiple_experiments(
        targetPaths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, targetPaths)

    print("Program completed")


if __name__ == "__main__":
    main()