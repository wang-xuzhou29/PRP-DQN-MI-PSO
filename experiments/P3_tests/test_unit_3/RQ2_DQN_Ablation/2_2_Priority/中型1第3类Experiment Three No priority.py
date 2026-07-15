import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# ==================== 编排规则专用范围 ====================
PATH_DEPTH_MIN, PATH_DEPTH_MAX = 1, 15
FILE_COUNT_MIN, FILE_COUNT_MAX = 1, 110000
ACCESS_LEVEL_MIN, ACCESS_LEVEL_MAX = 1, 4

# === 归一化/反归一化（全部整数）===
def normalize_state(state):
    pd, fc, al = state
    norm_pd = (pd - PATH_DEPTH_MIN) / (PATH_DEPTH_MAX - PATH_DEPTH_MIN)
    norm_fc = (fc - FILE_COUNT_MIN) / (FILE_COUNT_MAX - FILE_COUNT_MIN)
    norm_al = (al - ACCESS_LEVEL_MIN) / (ACCESS_LEVEL_MAX - ACCESS_LEVEL_MIN)
    return (norm_pd, norm_fc, norm_al)

def denormalize_state(norm_state):
    n_pd, n_fc, n_al = norm_state
    pd = int(round(n_pd * (PATH_DEPTH_MAX - PATH_DEPTH_MIN) + PATH_DEPTH_MIN))
    fc = int(round(n_fc * (FILE_COUNT_MAX - FILE_COUNT_MIN) + FILE_COUNT_MIN))
    al = int(round(n_al * (ACCESS_LEVEL_MAX - ACCESS_LEVEL_MIN) + ACCESS_LEVEL_MIN))
    pd = max(PATH_DEPTH_MIN, min(PATH_DEPTH_MAX, pd))
    fc = max(FILE_COUNT_MIN, min(FILE_COUNT_MAX, fc))
    al = max(ACCESS_LEVEL_MIN, min(ACCESS_LEVEL_MAX, al))
    return (pd, fc, al)

def is_valid_state(state):
    pd, fc, al = state
    return (PATH_DEPTH_MIN <= pd <= PATH_DEPTH_MAX and
            FILE_COUNT_MIN <= fc <= FILE_COUNT_MAX and
            ACCESS_LEVEL_MIN <= al <= ACCESS_LEVEL_MAX)

def clip_state(state):
    pd, fc, al = state
    pd = max(PATH_DEPTH_MIN, min(PATH_DEPTH_MAX, int(round(pd))))
    fc = max(FILE_COUNT_MIN, min(FILE_COUNT_MAX, int(round(fc))))
    al = max(ACCESS_LEVEL_MIN, min(ACCESS_LEVEL_MAX, int(round(al))))
    return (pd, fc, al)

def random_state():
    pd = random.randint(PATH_DEPTH_MIN, PATH_DEPTH_MAX)
    fc = random.randint(FILE_COUNT_MIN, FILE_COUNT_MAX)
    al = random.randint(ACCESS_LEVEL_MIN, ACCESS_LEVEL_MAX)
    return (pd, fc, al)

# === Metric Collector（保持不变）===
class EnhancedStandardMetricsCollector:
    def __init__(self, experiment_name="Enhanced_Standard_DQN"):
        self.experiment_name = experiment_name
        self.start_time = None
        self.end_time = None
        self.total_reward = 0
        self.td_errors = []
        self.final_output_similarities = []
        self.action_improvements = []
        self.total_memory_usage = 0
        self.memory_check_count = 0
        self.step_count = 0
        self.episode_rewards = []
        self.episode_similarities = []
        self.episode_td_errors = []
        self.episode_epsilon_values = []
        self.episode_memory_usage = []
        self.similar_paths_performance = []
        self.isolated_paths_performance = []
        self.milestone_data = {}
        self.convergence_window = 20
        self.convergence_threshold = 0.02
        self.convergence_detected_episode = None
        self.sample_efficiency_data = []
        self.performance_milestones = [0.6, 0.7, 0.75, 0.8]
        self.learning_curve_characteristics = {}
        self.early_vs_late_performance = {}
        self.xyz_coordinate_stats = {
            'x_values': [], 'y_values': [], 'z_values': [],
            'x_distribution': {}, 'y_distribution': {}, 'z_distribution': {}
        }

    def reset(self):
        self.__init__(self.experiment_name)

    def start_training(self):
        self.start_time = time.time()

    def end_training(self):
        self.end_time = time.time()

    def record_xyz_coordinates(self, x_val, y_val, z_val):
        self.xyz_coordinate_stats['x_values'].append(x_val)
        self.xyz_coordinate_stats['y_values'].append(y_val)
        self.xyz_coordinate_stats['z_values'].append(z_val)
        self.xyz_coordinate_stats['x_distribution'][x_val] = self.xyz_coordinate_stats['x_distribution'].get(x_val, 0) + 1
        self.xyz_coordinate_stats['y_distribution'][y_val] = self.xyz_coordinate_stats['y_distribution'].get(y_val, 0) + 1
        self.xyz_coordinate_stats['z_distribution'][z_val] = self.xyz_coordinate_stats['z_distribution'].get(z_val, 0) + 1

    def record_step_metrics(self, reward, td_error, triggered, target_path, x_coord=None, y_coord=None, z_coord=None):
        self.step_count += 1
        self.total_reward += reward
        self.td_errors.append(td_error)
        if x_coord is not None and y_coord is not None and z_coord is not None:
            self.record_xyz_coordinates(x_coord, y_coord, z_coord)
        process = psutil.Process(os.getpid())
        current_memory = process.memory_info().rss / 1024 / 1024
        self.total_memory_usage += current_memory
        self.memory_check_count += 1

    def record_episode_metrics(self, episode, episode_reward, avg_similarity, avg_td_error, epsilon,
                               path_group="similar"):
        self.episode_rewards.append(episode_reward)
        self.episode_similarities.append(avg_similarity)
        self.episode_td_errors.append(avg_td_error)
        self.episode_epsilon_values.append(epsilon)
        process = psutil.Process(os.getpid())
        current_memory = process.memory_info().rss / 1024 / 1024
        self.episode_memory_usage.append(current_memory)
        if path_group == "similar":
            self.similar_paths_performance.append({
                'episode': episode,
                'reward': episode_reward,
                'similarity': avg_similarity,
                'td_error': avg_td_error
            })
        else:
            self.isolated_paths_performance.append({
                'episode': episode,
                'reward': episode_reward,
                'similarity': avg_similarity,
                'td_error': avg_td_error
            })
        if episode in [50, 100, 150, 200, 250, 300, 400, 450, 500]:
            self.milestone_data[episode] = {
                'avg_reward': np.mean(self.episode_rewards[-10:]) if len(self.episode_rewards) >= 10 else episode_reward,
                'avg_similarity': avg_similarity,
                'avg_td_error': avg_td_error,
                'epsilon': epsilon,
                'memory_usage': current_memory,
                'total_steps': self.step_count
            }
        self._check_convergence(episode)
        self._check_performance_milestones(episode, avg_similarity)

    def _check_convergence(self, episode):
        if len(self.episode_similarities) >= self.convergence_window and self.convergence_detected_episode is None:
            recent = self.episode_similarities[-self.convergence_window:]
            if np.std(recent) < self.convergence_threshold:
                self.convergence_detected_episode = episode

    def _check_performance_milestones(self, episode, similarity):
        for milestone in self.performance_milestones:
            if similarity >= milestone and not any(data[1] == milestone for data in self.sample_efficiency_data):
                self.sample_efficiency_data.append((episode, milestone, self.step_count))

    def record_final_output_sample(self, triggered, target_path):
        if len(triggered | target_path) > 0:
            similarity = len(triggered & target_path) / len(triggered | target_path)
        else:
            similarity = 0.0
        self.final_output_similarities.append(similarity)

    def record_action_improvement(self, current_reward, prev_reward):
        if prev_reward is not None:
            improvement = current_reward - prev_reward
            self.action_improvements.append(1 if improvement > 0 else 0)

enhanced_standard_metrics = EnhancedStandardMetricsCollector("Enhanced_Standard_DQN_No_Priority")

# === 奖励函数 ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward

# ==================== 执行编排规则函数 ====================
def execute_orchestration_rules(a):
    path_depth, file_count, access_level = float(a[0]), int(a[1]), float(a[2])
    triggered = set()
    b = {}

    # 原始编排规则（与用户提供一致）
    if (path_depth >= 10) != (path_depth >= 12):
        b[0] = 1; triggered.add(1)
    if (path_depth >= 10) != (path_depth == 10):
        b[1] = 2; triggered.add(2)
    if (path_depth >= 10) != (path_depth >= 8):
        b[2] = 3; triggered.add(3)

    if path_depth >= 10:
        if (file_count >= 10000 and access_level >= 3) != (file_count >= 10000 and access_level >= 4.3):
            b[3] = 4; triggered.add(4)
        if (file_count >= 10000 and access_level >= 3) != (file_count >= 10000 and access_level == 3):
            b[4] = 5; triggered.add(5)
        if (file_count >= 10000 and access_level >= 3) != (file_count >= 10000 and access_level >= 3.2):
            b[5] = 6; triggered.add(6)
        if (file_count >= 10000 and access_level >= 3) != (file_count == 10000 and access_level >= 3):
            b[6] = 7; triggered.add(7)
        if (file_count >= 10000 and access_level >= 3) != (file_count >= 10000 and access_level >= 4):
            b[7] = 8; triggered.add(8)
        if (file_count >= 10000 and access_level >= 3) != (file_count >= 13400 and access_level >= 3.9):
            b[8] = 9; triggered.add(9)
        if (file_count >= 10000 and access_level >= 3) != (file_count != 10000 and access_level >= 3):
            b[9] = 10; triggered.add(10)
        if (file_count >= 10000 and access_level >= 3) != (file_count >= 10000 and access_level >= 5):
            b[10] = 11; triggered.add(11)

        if file_count >= 10000 and access_level >= 3:
            if (access_level == 3 and file_count >= 50000) != (access_level != 3 and file_count >= 50000):
                b[11] = 12; triggered.add(12)
            if (access_level == 3 and file_count >= 50000) != (access_level >= 3 and file_count >= 50000):
                b[12] = 13; triggered.add(13)
            if (access_level == 3 and file_count >= 50000) != (access_level <= 3 and file_count >= 50000):
                b[13] = 14; triggered.add(14)
            if (access_level == 3 and file_count >= 50000) != (access_level == 3 or file_count >= 50000):
                b[14] = 15; triggered.add(15)

            if (file_count >= 25000) != (file_count != 25000):
                b[15] = 16; triggered.add(16)
            if (file_count >= 25000) != (file_count >= 30000):
                b[16] = 17; triggered.add(17)
            if (file_count >= 25000) != (file_count >= 29000):
                b[17] = 18; triggered.add(18)

        if (file_count >= 5000 and access_level >= 2) != (file_count >= 5000 and access_level >= 2.5):
            b[18] = 19; triggered.add(19)
        if (file_count >= 5000 and access_level >= 2) != (file_count >= 5000 or access_level >= 3.5):
            b[19] = 20; triggered.add(20)
        if (file_count >= 5000 and access_level >= 2) != (file_count >= 5000 and access_level != 2):
            b[20] = 21; triggered.add(21)
        if (file_count >= 5000 and access_level >= 2) != (file_count != 5000 and access_level >= 2):
            b[21] = 22; triggered.add(22)
        if (file_count >= 5000 and access_level >= 2) != (file_count >= 5000 and access_level >= 3):
            b[22] = 23; triggered.add(23)
        if (file_count >= 5000 and access_level >= 2) != (file_count >= 5000 and access_level >= 3.3):
            b[23] = 24; triggered.add(24)

        if (file_count >= 1000) != (file_count >= 43000):
            b[24] = 25; triggered.add(25)
        if (file_count >= 1000) != (file_count >= 25000):
            b[25] = 26; triggered.add(26)
        if (file_count >= 1000) != (file_count >= 50000):
            b[26] = 27; triggered.add(27)

    if (path_depth >= 6) != (path_depth >= 7):
        b[27] = 28; triggered.add(28)
    if (path_depth >= 6) != (path_depth != 6):
        b[28] = 29; triggered.add(29)
    if (path_depth >= 6) != (path_depth >= 7):
        b[29] = 30; triggered.add(30)

    elif path_depth >= 6:
        if (file_count >= 20000 and access_level >= 3) != (file_count >= 20000 and access_level >= 4):
            b[30] = 31; triggered.add(31)
        if (file_count >= 20000 and access_level >= 3) != (file_count >= 20000 and access_level != 3):
            b[31] = 32; triggered.add(32)
        if (file_count >= 20000 and access_level >= 3) != (file_count >= 20000 and access_level == 3):
            b[32] = 33; triggered.add(33)
        if (file_count >= 20000 and access_level >= 3) != (file_count == 20000 and access_level >= 3):
            b[33] = 34; triggered.add(34)
        if (file_count >= 20000 and access_level >= 3) != (file_count != 20000 and access_level >= 3):
            b[34] = 35; triggered.add(35)

        if (file_count >= 8000 and access_level >= 2) != (file_count >= 54000 and access_level >= 2):
            b[35] = 36; triggered.add(36)
        if (file_count >= 8000 and access_level >= 2) != (file_count != 8000 and access_level >= 2):
            b[36] = 37; triggered.add(37)
        if (file_count >= 8000 and access_level >= 2) != (file_count >= 45000 and access_level >= 2):
            b[37] = 38; triggered.add(38)
        if (file_count >= 8000 and access_level >= 2) != (file_count >= 8000 and access_level != 2):
            b[38] = 39; triggered.add(39)
        if (file_count >= 8000 and access_level >= 2) != (file_count >= 8000 and access_level >= 3.3):
            b[39] = 40; triggered.add(40)
        if (file_count >= 8000 and access_level >= 2) != (file_count >= 8000 and access_level >= 2.5):
            b[40] = 41; triggered.add(41)
        if (file_count >= 8000 and access_level >= 2) != (file_count >= 80500 and access_level >= 2):
            b[41] = 42; triggered.add(42)

        elif file_count >= 8000 and access_level >= 2:
            if (access_level == 3) != (access_level <= 3):
                b[42] = 43; triggered.add(43)
            if (access_level == 3) != (access_level >= 3):
                b[43] = 44; triggered.add(44)

        if (file_count >= 2000) != (file_count != 2000):
            b[44] = 45; triggered.add(45)
        if (file_count >= 2000) != (file_count >= 62000):
            b[45] = 46; triggered.add(46)

    if (path_depth >= 3) != (path_depth != 3):
        b[46] = 47; triggered.add(47)
    if (path_depth >= 3) != (path_depth >= 3.5):
        b[47] = 48; triggered.add(48)

    elif path_depth >= 3:
        if (file_count >= 15000 and access_level >= 2) != (file_count != 15000 and access_level >= 2):
            b[48] = 49; triggered.add(49)
        if (file_count >= 15000 and access_level >= 2) != (file_count == 15000 and access_level >= 2):
            b[49] = 50; triggered.add(50)
        if (file_count >= 15000 and access_level >= 2) != (file_count >= 15000 and access_level != 2):
            b[50] = 51; triggered.add(51)
        if (file_count >= 15000 and access_level >= 2) != (file_count >= 15000 and access_level == 2):
            b[51] = 52; triggered.add(52)
        if (file_count >= 15000 and access_level >= 2) != (file_count >= 15000 and access_level >= 2.5):
            b[52] = 53; triggered.add(53)
        if (file_count >= 15000 and access_level >= 2) != (file_count >= 17500 and access_level >= 2):
            b[53] = 54; triggered.add(54)
        if (file_count >= 15000 and access_level >= 2) != (file_count >= 33330 and access_level >= 2):
            b[54] = 55; triggered.add(55)

        if file_count >= 15000 and access_level >= 2:
            if (access_level == 3) != (access_level >= 3):
                b[55] = 56; triggered.add(56)
            if (access_level == 3) != (access_level <= 3):
                b[56] = 57; triggered.add(57)

        if (file_count >= 5000) != (file_count != 5000):
            b[57] = 58; triggered.add(58)
        if (file_count >= 5000) != (file_count >= 55000):
            b[58] = 59; triggered.add(59)

    # 文件数量维度处理
    if (file_count >= 100000) != (file_count >= 90000):
        b[59] = 60; triggered.add(60)
    if (file_count >= 100000) != (file_count == 100000):
        b[60] = 61; triggered.add(61)

    if file_count >= 100000:
        if (path_depth >= 8 and access_level >= 3) != (path_depth >= 8 and access_level != 3):
            b[61] = 62; triggered.add(62)
        if (path_depth >= 8 and access_level >= 3) != (path_depth >= 8 and access_level == 3):
            b[62] = 63; triggered.add(63)
        if (path_depth >= 8 and access_level >= 3) != (path_depth == 8 and access_level >= 3):
            b[63] = 64; triggered.add(64)
        if (path_depth >= 8 and access_level >= 3) != (path_depth != 8 and access_level >= 3):
            b[64] = 65; triggered.add(65)
        if (path_depth >= 8 and access_level >= 3) != (path_depth >= 9 and access_level >= 3):
            b[65] = 66; triggered.add(66)
        if (path_depth >= 8 and access_level >= 3) != (path_depth >= 11 and access_level >= 3):
            b[66] = 67; triggered.add(67)
        if (path_depth >= 8 and access_level >= 3) != (path_depth >= 15 and access_level >= 3):
            b[67] = 68; triggered.add(68)

        if (path_depth >= 5 and access_level >= 2) != (path_depth != 5 and access_level >= 2):
            b[68] = 69; triggered.add(69)
        if (path_depth >= 5 and access_level >= 2) != (path_depth == 5 and access_level >= 2):
            b[69] = 70; triggered.add(70)
        if (path_depth >= 5 and access_level >= 2) != (path_depth >= 5 and access_level >= 2.9):
            b[70] = 71; triggered.add(71)
        if (path_depth >= 5 and access_level >= 2) != (path_depth >= 5 and access_level != 2):
            b[71] = 72; triggered.add(72)
        if (path_depth >= 5 and access_level >= 2) != (path_depth >= 5 and access_level == 2):
            b[72] = 73; triggered.add(73)
        if (path_depth >= 5 and access_level >= 2) != (path_depth >= 5 and access_level >= 2.5):
            b[73] = 74; triggered.add(74)
        if (path_depth >= 5 and access_level >= 2) != (path_depth >= 7 and access_level >= 2):
            b[74] = 75; triggered.add(75)

    if (file_count >= 50000) != (file_count != 50000):
        b[75] = 76; triggered.add(76)
    if (file_count >= 50000) != (file_count == 50000):
        b[76] = 77; triggered.add(77)

    elif file_count >= 50000:
        if (access_level >= 3) != (access_level != 3):
            b[77] = 78; triggered.add(78)
        if (access_level >= 3) != (access_level == 3):
            b[78] = 79; triggered.add(79)

    if (file_count >= 10000) != (file_count != 10000):
        b[79] = 80; triggered.add(80)
    if (file_count >= 10000) != (file_count == 10000):
        b[80] = 81; triggered.add(81)

    # 访问级别维度的扫描序列规划
    if (access_level == 3) != (access_level <= 3):
        b[81] = 82; triggered.add(82)
    if (access_level == 3) != (access_level >= 3):
        b[82] = 83; triggered.add(83)

    if access_level == 3:
        if (path_depth >= 8 and file_count >= 20000) != (path_depth != 8 and file_count >= 20000):
            b[83] = 84; triggered.add(84)
        if (path_depth >= 8 and file_count >= 20000) != (path_depth == 8 and file_count >= 20000):
            b[84] = 85; triggered.add(85)
        if (path_depth >= 8 and file_count >= 20000) != (path_depth >= 7 and file_count >= 20000):
            b[85] = 86; triggered.add(86)
        if (path_depth >= 8 and file_count >= 20000) != (path_depth >= 8 and file_count != 20000):
            b[86] = 87; triggered.add(87)
        if (path_depth >= 8 and file_count >= 20000) != (path_depth >= 8 and file_count == 20000):
            b[87] = 88; triggered.add(88)
        if (path_depth >= 8 and file_count >= 20000) != (path_depth >= 11 and file_count == 20000):
            b[88] = 89; triggered.add(89)
        if (path_depth >= 8 and file_count >= 20000) != (path_depth >= 8 and file_count == 25500):
            b[89] = 90; triggered.add(90)

        if (path_depth >= 5) != (path_depth != 5):
            b[90] = 91; triggered.add(91)
        if (path_depth >= 5) != (path_depth >= 3.5):
            b[91] = 92; triggered.add(92)
        if (path_depth >= 5) != (path_depth >= 6):
            b[92] = 93; triggered.add(93)

    if (access_level == 2) != (access_level >= 2):
        b[93] = 94; triggered.add(94)
    if (access_level == 2) != (access_level <= 2):
        b[94] = 95; triggered.add(95)

    return triggered

# 将规则函数赋给 execute_Tr
execute_Tr = execute_orchestration_rules

def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0

# ==================== 目标路径组（编排规则18条）====================
targetPaths = [
    # A1
    {1, 4, 6, 7, 8, 9, 11, 15, 17, 24, 25, 27, 31, 32, 34, 36, 38, 40, 42, 46, 50, 52, 55, 59, 62, 64, 67, 68, 70, 73, 76, 78, 81, 85, 88, 89, 90, 94},
    # A2
    {1, 2, 4, 6, 7, 8, 9, 11, 15, 17, 24, 25, 27, 31, 32, 34, 36, 38, 40, 42, 46, 50, 52, 55, 59, 62, 64, 68, 70, 73, 76, 78, 81, 85, 88, 89, 90, 94},
    # A3
    {3, 4, 5, 7, 11, 17, 25, 27, 33, 34, 36, 38, 42, 44, 46, 50, 52, 55, 56, 59, 63, 65, 66, 67, 68, 70, 73, 76, 79, 81, 83, 84, 88, 89, 90, 94},
    # A4
    {17, 19, 21, 23, 24, 25, 27, 28, 29, 30, 32, 36, 38, 39, 40, 41, 42, 43, 46, 50, 51, 53, 55, 57, 59, 70, 71, 72, 74, 75, 76, 78, 81, 82, 84},
    # A5
    {19, 21, 23, 24, 25, 27, 29, 32, 36, 38, 39, 40, 41, 42, 43, 46, 50, 51, 53, 55, 57, 59, 69, 71, 72, 74, 75, 76, 78, 81, 82, 84, 91, 93},
    # A6
    {16, 19, 21, 23, 24, 25, 26, 27, 29, 36, 38, 39, 40, 41, 42, 43, 46, 50, 51, 53, 55, 57, 59, 69, 71, 72, 74, 75, 76, 78, 81, 82, 91, 93},
    # A7
    {3, 4, 6, 7, 8, 9, 11, 15, 16, 24, 25, 26, 27, 35, 36, 38, 40, 42, 46, 50, 52, 54, 55, 59, 62, 64, 67, 68, 70, 73, 76, 78, 81, 87, 94},
    # A8
    {1, 12, 14, 15, 19, 21, 23, 24, 32, 39, 40, 41, 42, 43, 46, 50, 51, 53, 57, 59, 62, 70, 71, 72, 74, 77, 78, 81, 82, 85, 88, 89, 90},
    # A9
    {3, 4, 5, 7, 11, 12, 13, 15, 33, 34, 42, 44, 46, 50, 52, 56, 59, 63, 65, 66, 67, 68, 70, 73, 77, 79, 81, 83, 84, 88, 89, 90, 94},
    # A10
    {17, 19, 21, 23, 24, 25, 27, 32, 36, 38, 39, 40, 41, 42, 43, 46, 50, 51, 53, 55, 57, 59, 70, 71, 72, 74, 76, 78, 81, 82, 84, 86},
    # A11
    {17, 19, 21, 23, 24, 25, 27, 29, 32, 36, 38, 39, 40, 41, 42, 43, 46, 47, 48, 50, 51, 53, 55, 57, 59, 69, 76, 78, 81, 82, 84, 91},
    # A12
    {16, 19, 21, 23, 24, 25, 26, 27, 28, 29, 30, 36, 38, 39, 40, 41, 42, 43, 46, 49, 57, 59, 70, 71, 72, 74, 75, 76, 78, 80, 82},
    # A13
    {16, 19, 21, 23, 24, 25, 26, 27, 29, 36, 38, 39, 40, 41, 42, 43, 46, 50, 51, 53, 55, 57, 59, 69, 76, 78, 81, 82, 91, 92},
    # A14
    {1, 10, 16, 20, 22, 25, 26, 27, 35, 37, 44, 45, 49, 56, 58, 63, 64, 67, 68, 70, 73, 76, 79, 80, 83, 87, 94},
    # A15
    {1, 2, 17, 20, 21, 25, 27, 32, 39, 43, 46, 51, 57, 59, 62, 72, 76, 78, 81, 82, 85, 88, 89, 90, 95},
    # A16
    {4, 5, 7, 11, 17, 18, 25, 27, 29, 33, 34, 36, 38, 42, 44, 46, 50, 52, 55, 56, 59, 65, 69, 73, 75, 76, 79, 81, 83, 84, 91, 93, 94},
    # A17
    {4, 6, 7, 8, 9, 11, 12, 24, 31, 32, 34, 40, 50, 52, 61, 65, 70, 73, 77, 78, 81, 84, 86, 94},
    # A18
    {1, 2, 4, 6, 7, 8, 9, 11, 12, 24, 31, 32, 34, 40, 50, 52, 60, 62, 64, 68, 70, 73, 77, 78, 81, 85, 88, 89, 90, 94}
]

def compute_path_similarity_matrix(paths):
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix

def group_paths_by_similarity(paths):
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.mean(avg_sim_scores)
    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]
    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)
    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group

# === 标准经验回放 ===
class StandardExperienceReplay:
    def __init__(self, capacity=10000):
        self.capacity = capacity
        self.buffer = deque(maxlen=self.capacity)

    def append(self, experience):
        self.buffer.append(experience[:5])

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return [], [], None
        batch_indices = np.random.choice(len(self.buffer), batch_size, replace=False)
        batch = [self.buffer[idx] for idx in batch_indices]
        return batch, batch_indices, None

    def __len__(self):
        return len(self.buffer)

    def get_high_reward_samples(self, target_path, num_samples=20):
        if len(self.buffer) == 0:
            return []
        samples_with_recalculated_scores = []
        for experience in self.buffer:
            norm_state_tensor = experience[0]
            norm_state = tuple(norm_state_tensor.cpu().numpy().flatten())
            state_tuple = denormalize_state(norm_state)
            triggered = execute_Tr(state_tuple[0], state_tuple[1], state_tuple[2])
            new_reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_recalculated_scores.append((state_tuple, new_reward, sim, triggered))
        samples_with_recalculated_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_recalculated_scores[:num_samples]

def load_path_data(file_path):
    path_data = []
    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines[2:]:
            parts = line.strip().split("\t")
            values = parts[0].split()
            pd = int(values[0])
            fc = int(values[1])
            al = int(values[2])
            path_data.append((pd, fc, al))
    return path_data

# === DQN 网络 ===
class DQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, 128)
        self.fc2 = nn.Linear(128, 64)
        self.fc3 = nn.Linear(64, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)

# === 标准 DQN Agent ===
class StandardDQNAgent:
    def __init__(self, state_dim, action_dim, replay_buffer, gamma=0.99, epsilon=1.0,
                 epsilon_decay=0.995, epsilon_min=0.1, learning_rate=0.001):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        # 为每个维度定制步长（全部为整数）
        step_lists = [
            [-3, -2, -1, 0, 1, 2, 3, 4, 5, 6],           # path_depth
            [-5000, -2000, -1000, -500, -100, 0, 100, 500, 1000, 5000],  # file_count
            [-1, -1, -1, 0, 0, 0, 1, 1, 1, 2]            # access_level（重复以填满10个）
        ]
        dim = action_idx // 10
        idx = action_idx % 10
        if dim == 0:
            return (step_lists[0][idx], 0, 0)
        elif dim == 1:
            return (0, step_lists[1][idx], 0)
        else:
            return (0, 0, step_lists[2][idx])

    def act(self, state):
        if random.random() < self.epsilon:
            return random.randrange(self.action_dim)
        norm_state = normalize_state(state)
        state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)
        return torch.argmax(q_values, dim=1).item()

    def store_transition(self, state, action, reward, next_state, done):
        norm_state = normalize_state(state)
        norm_next_state = normalize_state(next_state)
        norm_state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
        norm_next_state_tensor = torch.tensor(norm_next_state, dtype=torch.float32).unsqueeze(0).to(device)

        with torch.no_grad():
            q_values = self.model(norm_state_tensor)
            next_q_values = self.target_model(norm_next_state_tensor)
            max_next_q_values = next_q_values.max(1)[0]
            target_q_values = reward + (self.gamma * max_next_q_values * (1 - done))
            td_error = torch.abs(q_values[0][action] - target_q_values).item()

        self.replay_buffer.append((norm_state_tensor, action, reward, norm_next_state_tensor, done))
        return td_error

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return
        batch, batch_indices, _ = self.replay_buffer.sample(batch_size)
        states, actions, rewards, next_states, dones = zip(*batch)

        states = torch.tensor(np.array([s.cpu().numpy().flatten() for s in states]), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array([ns.cpu().numpy().flatten() for ns in next_states]), dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))

        loss = nn.MSELoss()(current_q_values, target_q_values)
        self.optimizer.zero_grad()
        loss.backward()
        self.optimizer.step()

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())

# ==================== 样本生成（相似组）====================
def generate_samples_for_similar_paths(similar_group_indices, num_total=2000, top_k=200):
    def compute_robustness(state, path):
        base = execute_Tr(state[0], state[1], state[2])
        if not base: return 0.0
        rob, neighbors = 0.0, 0
        # 邻域步长：path_depth ±1, file_count ±500, access_level ±1
        for dp in [-1, 0, 1]:
            for df in [-500, 0, 500]:
                for da in [-1, 0, 1]:
                    if dp == df == da == 0:
                        continue
                    neighbor_state = (state[0] + dp, state[1] + df, state[2] + da)
                    if not is_valid_state(neighbor_state):
                        continue
                    neighbor = clip_state(neighbor_state)
                    n_trig = execute_Tr(neighbor[0], neighbor[1], neighbor[2])
                    if not n_trig:
                        continue
                    rob += jaccard_similarity(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"enhanced_standard_path{path_id}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Enhanced Standard Path {path_id}\n")
            f.write("path_depth file_count access_level\tScore\tSimilarity\tLengthDiff\tRobustness\n")
            for s in samples:
                pd, fc, al = s[0]
                f.write(f"{pd} {fc} {al}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_enhanced_standard"
    for path_idx in similar_group_indices:
        path = targetPaths[path_idx]
        samples = []
        attempts = 0
        while len(samples) < top_k and attempts < num_total * 5:
            attempts += 1
            state = random_state()
            triggered = execute_Tr(state[0], state[1], state[2])
            if not triggered: continue
            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            score = 0.55 * sim + 0.25 * len_diff + 0.2 * rob
            samples.append((state, score, sim, len_diff, rob))
        if samples:
            samples.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples[:top_k], base_dir=base_dir)

# ==================== 训练相似组 ====================
def enhanced_standard_generate_and_train_for_similar_paths(agent, similar_group, path_documents, episodes=500,
                                                           batch_size=32, steps_per_test=5, replay_times=10,
                                                           is_isolated=False):
    trained_paths = set()
    update_target_every = 100
    global_steps = 0

    for episode in range(episodes):
        episode_reward = 0
        episode_similarities = []
        episode_td_errors = []

        for path_idx in similar_group:
            if path_idx in trained_paths:
                continue

            file_path = os.path.join(path_documents,
                                     f"enhanced_standard_path{path_idx + 1}{'_isolated' if is_isolated else ''}.txt")
            path_data = load_path_data(file_path)
            target_path = targetPaths[path_idx]

            BATCH_SIZE = 50
            N_SAMPLES = 200
            N_STEPS = 10
            REPLAY_TIMES = 3

            for batch_start in range(0, N_SAMPLES, BATCH_SIZE):
                batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)
                for test_data in range(batch_start, batch_end):
                    if test_data >= len(path_data):
                        break

                    step_count = 0
                    state = path_data[test_data]
                    prev_state = None
                    prev_triggered = None
                    prev_reward = None

                    for step in range(N_STEPS):
                        legal_actions = []
                        for a in range(agent.action_dim):
                            dx, dy, dz = agent.decode_action(a)
                            cand_next = (state[0] + dx, state[1] + dy, state[2] + dz)
                            if is_valid_state(cand_next):
                                legal_actions.append(a)
                        if not legal_actions:
                            break

                        if random.random() < agent.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            norm_state = normalize_state(state)
                            state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        dx, dy, dz = agent.decode_action(action)
                        next_state = clip_state((state[0] + dx, state[1] + dy, state[2] + dz))

                        triggered = execute_Tr(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                        done = (step_count == N_STEPS - 1)

                        td_error = agent.store_transition(state, action, reward, next_state, done)
                        enhanced_standard_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                      next_state[0], next_state[1], next_state[2])

                        episode_similarities.append(jaccard_similarity(triggered, target_path))
                        episode_td_errors.append(td_error)

                        if prev_reward is not None:
                            enhanced_standard_metrics.record_action_improvement(reward, prev_reward)

                        prev_state = state
                        prev_triggered = triggered
                        prev_reward = reward
                        state = next_state
                        step_count += 1
                        episode_reward += reward
                        global_steps += 1

                        if global_steps % update_target_every == 0:
                            agent.update_target_model()

                for _ in range(REPLAY_TIMES):
                    if len(agent.replay_buffer) >= batch_size:
                        agent.train(batch_size)

            trained_paths.add(path_idx)

        avg_similarity = np.mean(episode_similarities) if episode_similarities else 0
        avg_td_error = np.mean(episode_td_errors) if episode_td_errors else 0
        enhanced_standard_metrics.record_episode_metrics(episode, episode_reward, avg_similarity, avg_td_error,
                                                         agent.epsilon, "similar")

        if episode % 10 == 0:
            agent.update_target_model()

        if len(trained_paths) == len(similar_group):
            break

    return agent

# ==================== 生成孤立组样本（利用相似组模型）====================
def generate_samples_for_isolated_paths_standard(agent_similar, isolated_group_indices, num_total=2000, top_k=200):
    def compute_q_value(state, agent):
        norm_state = normalize_state(state)
        state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = agent.model(state_tensor)
        return q_values.max().item()

    def compute_robustness(state, path):
        base = execute_Tr(state[0], state[1], state[2])
        if not base: return 0.0
        rob, neighbors = 0.0, 0
        for dp in [-1, 0, 1]:
            for df in [-500, 0, 500]:
                for da in [-1, 0, 1]:
                    if dp == df == da == 0: continue
                    neighbor_state = (state[0] + dp, state[1] + df, state[2] + da)
                    if not is_valid_state(neighbor_state): continue
                    neighbor = clip_state(neighbor_state)
                    n_trig = execute_Tr(neighbor[0], neighbor[1], neighbor[2])
                    if not n_trig: continue
                    rob += jaccard_similarity(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"enhanced_standard_path{path_id}_isolated.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Enhanced Standard Isolated Path {path_id}\n")
            f.write("path_depth file_count access_level\tScore\tSimilarity\tLengthDiff\tRobustness\tQ_complement\n")
            for s in samples:
                pd, fc, al = s[0]
                f.write(f"{pd} {fc} {al}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\t{s[5]:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_enhanced_standard"
    for path_idx in isolated_group_indices:
        path = targetPaths[path_idx]
        samples_raw = []
        attempts = 0

        print(f"Path  {path_idx + 1} ...")
        while len(samples_raw) < num_total and attempts < num_total * 5:
            attempts += 1
            state = random_state()
            triggered = execute_Tr(state[0], state[1], state[2])
            if not triggered: continue
            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            q_value = compute_q_value(state, agent_similar)
            samples_raw.append((state, sim, len_diff, rob, q_value))

        if not samples_raw:
            print(f"Path  {path_idx + 1}: no samples")
            continue

        q_values_list = [s[4] for s in samples_raw]
        q_min = min(q_values_list)
        q_max = max(q_values_list)

        print(f"Path  {path_idx + 1}: generated {len(samples_raw)} candidates")
        print(f"  Q range: [{q_min:.4f}, {q_max:.4f}]")

        samples_final = []
        for state, sim, len_diff, rob, q_value in samples_raw:
            if q_max - q_min > 1e-6:
                q_normalized = (q_value - q_min) / (q_max - q_min)
            else:
                q_normalized = 0.5
            q_complement = 1.0 - q_normalized
            score = 0.28 * sim + 0.1 * len_diff + 0.19 * rob + 0.43 * q_complement
            samples_final.append((state, score, sim, len_diff, rob, q_complement))

        if samples_final:
            samples_final.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples_final[:top_k], base_dir=base_dir)
            print(f"Path  {path_idx + 1}: saved {min(top_k, len(samples_final))} samples")
            print(f"  Q_complement range: [{samples_final[-1][5]:.4f}, {samples_final[0][5]:.4f}]")

# ==================== 训练孤立组（两阶段）====================
def enhanced_standard_generate_and_train_for_isolated_paths(agent_similar, agent_isolated, similar_group,
                                                            isolated_group, path_documents, episodes=500, batch_size=32,
                                                            is_isolated=True):
    trained_paths = set()
    update_target_every = 100
    global_steps = 0

    stage1_samples_pool = {}
    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_similar.replay_buffer.get_high_reward_samples(target_path, num_samples=100)
        stage1_samples_pool[path_idx] = high_reward_samples

    for episode in range(episodes):
        episode_reward = 0
        episode_similarities = []
        episode_td_errors = []

        for path_idx in isolated_group:
            if path_idx in trained_paths:
                continue

            file_path = os.path.join(path_documents, f"enhanced_standard_path{path_idx + 1}_isolated.txt")
            stage2_path_data = load_path_data(file_path)
            stage1_samples = stage1_samples_pool.get(path_idx, [])
            target_path = targetPaths[path_idx]

            BATCH_SIZE = 50
            N_SAMPLES_STAGE2 = min(140, len(stage2_path_data))
            N_SAMPLES_STAGE1 = min(60, len(stage1_samples))
            N_STEPS = 10
            REPLAY_TIMES = 3

            for batch_start in range(0, N_SAMPLES_STAGE2, BATCH_SIZE):
                batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES_STAGE2)
                for test_data_idx in range(batch_start, batch_end):
                    if test_data_idx >= len(stage2_path_data):
                        break

                    step_count = 0
                    state = stage2_path_data[test_data_idx]
                    prev_state = None
                    prev_triggered = None
                    prev_reward = None

                    for step in range(N_STEPS):
                        legal_actions = []
                        for a in range(agent_isolated.action_dim):
                            dx, dy, dz = agent_isolated.decode_action(a)
                            cand_next = (state[0] + dx, state[1] + dy, state[2] + dz)
                            if is_valid_state(cand_next):
                                legal_actions.append(a)

                        if not legal_actions:
                            break

                        if random.random() < agent_isolated.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            norm_state = normalize_state(state)
                            state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent_isolated.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        dx, dy, dz = agent_isolated.decode_action(action)
                        next_state = clip_state((state[0] + dx, state[1] + dy, state[2] + dz))

                        triggered = execute_Tr(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                        if target_path.issubset(triggered):
                            reward += 2.0

                        done = (step_count == N_STEPS - 1)

                        td_error = agent_isolated.store_transition(state, action, reward, next_state, done)
                        enhanced_standard_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                      next_state[0], next_state[1], next_state[2])

                        episode_similarities.append(jaccard_similarity(triggered, target_path))
                        episode_td_errors.append(td_error)

                        if prev_reward is not None:
                            enhanced_standard_metrics.record_action_improvement(reward, prev_reward)

                        prev_state = state
                        prev_triggered = triggered
                        prev_reward = reward
                        state = next_state
                        step_count += 1
                        episode_reward += reward
                        global_steps += 1

                        if global_steps % update_target_every == 0:
                            agent_isolated.update_target_model()

            if stage1_samples:
                for sample_idx in range(N_SAMPLES_STAGE1):
                    if sample_idx >= len(stage1_samples):
                        break

                    stage1_state_tuple, _, _, _ = stage1_samples[sample_idx]
                    step_count = 0
                    state = stage1_state_tuple
                    prev_state = None
                    prev_triggered = None
                    prev_reward = None

                    for step in range(N_STEPS):
                        legal_actions = []
                        for a in range(agent_isolated.action_dim):
                            dx, dy, dz = agent_isolated.decode_action(a)
                            cand_next = (state[0] + dx, state[1] + dy, state[2] + dz)
                            if is_valid_state(cand_next):
                                legal_actions.append(a)

                        if not legal_actions:
                            break

                        if random.random() < agent_isolated.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            norm_state = normalize_state(state)
                            state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent_isolated.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        dx, dy, dz = agent_isolated.decode_action(action)
                        next_state = clip_state((state[0] + dx, state[1] + dy, state[2] + dz))

                        triggered = execute_Tr(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                        reward *= 0.8

                        done = (step_count == N_STEPS - 1)

                        td_error = agent_isolated.store_transition(state, action, reward, next_state, done)
                        enhanced_standard_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                      next_state[0], next_state[1], next_state[2])

                        episode_similarities.append(jaccard_similarity(triggered, target_path))
                        episode_td_errors.append(td_error)

                        if prev_reward is not None:
                            enhanced_standard_metrics.record_action_improvement(reward, prev_reward)

                        prev_state = state
                        prev_triggered = triggered
                        prev_reward = reward
                        state = next_state
                        step_count += 1
                        episode_reward += reward
                        global_steps += 1

                        if global_steps % update_target_every == 0:
                            agent_isolated.update_target_model()

            for replay_round in range(REPLAY_TIMES):
                if len(agent_isolated.replay_buffer) >= batch_size:
                    agent_isolated.train(batch_size)

            trained_paths.add(path_idx)

        avg_similarity = np.mean(episode_similarities) if episode_similarities else 0
        avg_td_error = np.mean(episode_td_errors) if episode_td_errors else 0
        enhanced_standard_metrics.record_episode_metrics(episode, episode_reward, avg_similarity, avg_td_error,
                                                         agent_isolated.epsilon, "isolated")

        if episode % 10 == 0:
            agent_isolated.update_target_model()

        if len(trained_paths) == len(isolated_group):
            break

    return agent_isolated

# ==================== 导出 Excel（双sheet）====================
def append_metrics_to_combined_excel(metrics_collector, agent_similar, agent_isolated, similar_group, isolated_group,
                                     targetPaths, filepath, run_number):
    training_time = metrics_collector.end_time - metrics_collector.start_time if metrics_collector.end_time else 0
    avg_memory = metrics_collector.total_memory_usage / metrics_collector.memory_check_count if metrics_collector.memory_check_count > 0 else 0
    avg_similarity = np.mean(metrics_collector.final_output_similarities) if metrics_collector.final_output_similarities else 0
    avg_td_error = np.mean(metrics_collector.td_errors) if metrics_collector.td_errors else 0
    action_improve_rate = np.mean(metrics_collector.action_improvements) if metrics_collector.action_improvements else 0

    performance_row = {
        'Run': run_number,
        'Average Similarity': f"{avg_similarity:.4f}",
        'TD Error': f"{avg_td_error:.4f}",
        'Action Improve Rate': f"{action_improve_rate:.4f}",
        'Memory(MB)': f"{avg_memory:.2f}"
    }

    sample_rows = []

    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_similar.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, reward, sim, triggered in high_reward_samples:
            sample_rows.append({
                'Run': run_number,
                'Group Type': 'Similar',
                'Path ID': path_idx + 1,
                'Path_Depth': state_tuple[0],
                'File_Count': state_tuple[1],
                'Access_Level': state_tuple[2],
                'Similarity': f"{sim:.4f}",
                'Reward': f"{reward:.2f}",
                'Triggered Count': len(triggered),
                'Target Count': len(target_path),
                'Triggered Rules': str(sorted(triggered)),
                'Target Rules': str(sorted(target_path))
            })

    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, reward, sim, triggered in high_reward_samples:
            sample_rows.append({
                'Run': run_number,
                'Group Type': 'Isolated',
                'Path ID': path_idx + 1,
                'Path_Depth': state_tuple[0],
                'File_Count': state_tuple[1],
                'Access_Level': state_tuple[2],
                'Similarity': f"{sim:.4f}",
                'Reward': f"{reward:.2f}",
                'Triggered Count': len(triggered),
                'Target Count': len(target_path),
                'Triggered Rules': str(sorted(triggered)),
                'Target Rules': str(sorted(target_path))
            })

    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    if os.path.exists(filepath):
        try:
            df_performance = pd.read_excel(filepath, sheet_name='Metric')
            df_performance = pd.concat([df_performance, pd.DataFrame([performance_row])], ignore_index=True)
        except:
            df_performance = pd.DataFrame([performance_row])
        try:
            df_samples = pd.read_excel(filepath, sheet_name='final samples')
            df_samples = pd.concat([df_samples, pd.DataFrame(sample_rows)], ignore_index=True)
        except:
            df_samples = pd.DataFrame(sample_rows)
    else:
        df_performance = pd.DataFrame([performance_row])
        df_samples = pd.DataFrame(sample_rows)

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df_performance.to_excel(writer, sheet_name='Metric', index=False)
        df_samples.to_excel(writer, sheet_name='final samples', index=False)

        workbook = writer.book
        ws_performance = writer.sheets['Metric']
        ws_performance.column_dimensions['A'].width = 15
        for col in ['B', 'C', 'D', 'E']:
            ws_performance.column_dimensions[col].width = 20
        header_font = Font(bold=True, size=11)
        for cell in ws_performance[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in ws_performance.iter_rows(min_row=2, max_row=ws_performance.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        ws_samples = writer.sheets['final samples']
        column_widths = {
            'A': 12, 'B': 15, 'C': 12, 'D': 14, 'E': 14, 'F': 16,
            'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 40, 'L': 40
        }
        for col, width in column_widths.items():
            ws_samples.column_dimensions[col].width = width
        for cell in ws_samples[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in ws_samples.iter_rows(min_row=2, max_row=ws_samples.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    print(f"Run {run_number} results appended to: {filepath}")

# ==================== 单次实验主流程 ====================
def run_single_experiment(run_number, results_save_dir):
    print(f"\n{'=' * 80}")
    print(f"Start run {run_number}")
    print(f"{'=' * 80}\n")

    enhanced_standard_metrics.reset()
    enhanced_standard_metrics.start_training()

    model_path_similar = os.path.join(results_save_dir, f"run{run_number}_similar.pth")
    model_path_isolated = os.path.join(results_save_dir, f"run{run_number}_isolated.pth")
    path_documents = r"D:\Experiment\CNN\DQNNEW\path_samples_enhanced_standard"

    similar_group, isolated_group = group_paths_by_similarity(targetPaths)

    # 第一次运行时生成相似组样本
    if run_number == 1:
        generate_samples_for_similar_paths(similar_group, num_total=2000, top_k=200)

    replay_buffer = StandardExperienceReplay(capacity=10000)
    state_dim = 3
    action_dim = 30
    agent = StandardDQNAgent(state_dim, action_dim, replay_buffer)

    agent = enhanced_standard_generate_and_train_for_similar_paths(agent, similar_group, path_documents, episodes=500,
                                                                   batch_size=32, is_isolated=False)

    os.makedirs(os.path.dirname(model_path_similar), exist_ok=True)
    torch.save({
        'model_state_dict': agent.model.state_dict(),
        'optimizer_state_dict': agent.optimizer.state_dict(),
        'epsilon': agent.epsilon
    }, model_path_similar)

    # 第一次运行时生成孤立组样本
    if run_number == 1:
        generate_samples_for_isolated_paths_standard(agent, isolated_group, num_total=2000, top_k=200)

    isolated_replay_buffer = StandardExperienceReplay(capacity=15000)
    agent_isolated = StandardDQNAgent(state_dim, action_dim, isolated_replay_buffer)

    try:
        checkpoint = torch.load(model_path_similar)
        agent_isolated.model.load_state_dict(checkpoint['model_state_dict'])
        agent_isolated.target_model.load_state_dict(checkpoint['model_state_dict'])
        agent_isolated.epsilon = checkpoint.get('epsilon', 0.5)
    except Exception as e:
        pass

    agent_isolated = enhanced_standard_generate_and_train_for_isolated_paths(
        agent_similar=agent,
        agent_isolated=agent_isolated,
        similar_group=similar_group,
        isolated_group=isolated_group,
        path_documents=path_documents,
        episodes=500,
        batch_size=32,
        is_isolated=True
    )

    os.makedirs(os.path.dirname(model_path_isolated), exist_ok=True)
    torch.save({
        'model_state_dict': agent_isolated.model.state_dict(),
        'optimizer_state_dict': agent_isolated.optimizer.state_dict(),
        'epsilon': agent_isolated.epsilon
    }, model_path_isolated)

    enhanced_standard_metrics.end_training()

    # 记录 final samples 相似度
    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, _, sim, triggered in high_reward_samples:
            enhanced_standard_metrics.record_final_output_sample(triggered, target_path)

    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, _, sim, triggered in high_reward_samples:
            enhanced_standard_metrics.record_final_output_sample(triggered, target_path)

    combined_excel_path = os.path.join(results_save_dir, "__.xlsx")
    append_metrics_to_combined_excel(
        metrics_collector=enhanced_standard_metrics,
        agent_similar=agent,
        agent_isolated=agent_isolated,
        similar_group=similar_group,
        isolated_group=isolated_group,
        targetPaths=targetPaths,
        filepath=combined_excel_path,
        run_number=run_number
    )

    avg_similarity = np.mean(enhanced_standard_metrics.final_output_similarities)
    print(f"\nRun {run_number} completed:")
    print(f"  Average Similarity: {avg_similarity:.4f}")
    print(f"  Total steps: {enhanced_standard_metrics.step_count}")

# ==================== 主程序 ====================
if __name__ == "__main__":
    results_save_dir = r"D:\Experiment\CNN\_"   # 请根据需要修改
    os.makedirs(results_save_dir, exist_ok=True)

    NUM_RUNS = 20
    print("=" * 80)
    print(f"Running {NUM_RUNS} runs of Standard DQN (Orchestration Rules)")
    print(f"State ranges: path_depth[{PATH_DEPTH_MIN},{PATH_DEPTH_MAX}], "
          f"file_count[{FILE_COUNT_MIN},{FILE_COUNT_MAX}], "
          f"access_level[{ACCESS_LEVEL_MIN},{ACCESS_LEVEL_MAX}]")
    print("=" * 80)

    for run in range(1, NUM_RUNS + 1):
        try:
            run_single_experiment(run, results_save_dir)
        except Exception as e:
            print(f"\nRun {run} failed: {str(e)}")
            import traceback
            traceback.print_exc()
            continue

    print("\n" + "=" * 80)
    print(f"All {NUM_RUNS} runs completed.")
    print(f"Results saved to: {results_save_dir}")
    print("=" * 80)