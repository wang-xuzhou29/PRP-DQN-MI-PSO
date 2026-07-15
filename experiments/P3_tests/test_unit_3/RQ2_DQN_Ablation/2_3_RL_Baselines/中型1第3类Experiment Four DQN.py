import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
import numpy as np
import random
import time
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# ==================== 编排规则专用配置 ====================
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'MIN_X': 1, 'MAX_X': 15,          # path_depth
    'MIN_Y': 1, 'MAX_Y': 110000,      # file_count
    'MIN_Z': 1, 'MAX_Z': 4,           # access_level
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 5,
    'NUM_ROUNDS': 5,
    'REPLAY_BATCH_SIZE': 64,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'LEARNING_RATE': 3e-4,
    'NUM_RUNS': 20,
    'TOP_K_SAMPLES': 20,
    'REPLAY_BUFFER_CAPACITY': 20000,
    # 编排规则的目标路径（18条）
    'TARGET_PATHS': [
        # A1
        {1, 4, 6, 7, 8, 9, 11, 15, 17, 24, 25, 27, 31, 32, 34, 36, 38, 40, 42, 46, 50, 52, 55, 59, 62, 64, 67, 68, 70, 73,
         76, 78, 81, 85, 88, 89, 90, 94},
        # A2
        {1, 2, 4, 6, 7, 8, 9, 11, 15, 17, 24, 25, 27, 31, 32, 34, 36, 38, 40, 42, 46, 50, 52, 55, 59, 62, 64, 68, 70, 73,
         76, 78, 81, 85, 88, 89, 90, 94},
        # A3
        {3, 4, 5, 7, 11, 17, 25, 27, 33, 34, 36, 38, 42, 44, 46, 50, 52, 55, 56, 59, 63, 65, 66, 67, 68, 70, 73, 76, 79, 81,
         83, 84, 88, 89, 90, 94},
        # A4
        {17, 19, 21, 23, 24, 25, 27, 28, 29, 30, 32, 36, 38, 39, 40, 41, 42, 43, 46, 50, 51, 53, 55, 57, 59, 70, 71, 72, 74,
         75, 76, 78, 81, 82, 84},
        # A5
        {19, 21, 23, 24, 25, 27, 29, 32, 36, 38, 39, 40, 41, 42, 43, 46, 50, 51, 53, 55, 57, 59, 69, 71, 72, 74, 75, 76, 78,
         81, 82, 84, 91, 93},
        # A6
        {16, 19, 21, 23, 24, 25, 26, 27, 29, 36, 38, 39, 40, 41, 42, 43, 46, 50, 51, 53, 55, 57, 59, 69, 71, 72, 74, 75, 76,
         78, 81, 82, 91, 93},
        # A7
        {3, 4, 6, 7, 8, 9, 11, 15, 16, 24, 25, 26, 27, 35, 36, 38, 40, 42, 46, 50, 52, 54, 55, 59, 62, 64, 67, 68, 70, 73,
         76, 78, 81, 87, 94},
        # A8
        {1, 12, 14, 15, 19, 21, 23, 24, 32, 39, 40, 41, 42, 43, 46, 50, 51, 53, 57, 59, 62, 70, 71, 72, 74, 77, 78, 81, 82,
         85, 88, 89, 90},
        # A9
        {3, 4, 5, 7, 11, 12, 13, 15, 33, 34, 42, 44, 46, 50, 52, 56, 59, 63, 65, 66, 67, 68, 70, 73, 77, 79, 81, 83, 84, 88,
         89, 90, 94},
        # A10
        {17, 19, 21, 23, 24, 25, 27, 32, 36, 38, 39, 40, 41, 42, 43, 46, 50, 51, 53, 55, 57, 59, 70, 71, 72, 74, 76, 78, 81,
         82, 84, 86},
        # A11
        {17, 19, 21, 23, 24, 25, 27, 29, 32, 36, 38, 39, 40, 41, 42, 43, 46, 47, 48, 50, 51, 53, 55, 57, 59, 69, 76, 78, 81,
         82, 84, 91},
        # A12
        {16, 19, 21, 23, 24, 25, 26, 27, 28, 29, 30, 36, 38, 39, 40, 41, 42, 43, 46, 49, 57, 59, 70, 71, 72, 74, 75, 76, 78,
         80, 82},
        # A13
        {16, 19, 21, 23, 24, 25, 26, 27, 29, 36, 38, 39, 40, 41, 42, 43, 46, 50, 51, 53, 55, 57, 59, 69, 76, 78, 81, 82, 91,
         92},
        # A14
        {1, 10, 16, 20, 22, 25, 26, 27, 35, 37, 44, 45, 49, 56, 58, 63, 64, 67, 68, 70, 73, 76, 79, 80, 83, 87, 94},
        # A15
        {1, 2, 17, 20, 21, 25, 27, 32, 39, 43, 46, 51, 57, 59, 62, 72, 76, 78, 81, 82, 85, 88, 89, 90, 95},
        # A16
        {4, 5, 7, 11, 17, 18, 25, 27, 29, 33, 34, 36, 38, 42, 44, 46, 50, 52, 55, 56, 59, 65, 69, 73, 75, 76, 79, 81, 83,
         84, 91, 93, 94},
        # A17
        {4, 6, 7, 8, 9, 11, 12, 24, 31, 32, 34, 40, 50, 52, 61, 65, 70, 73, 77, 78, 81, 84, 86, 94},
        # A18
        {1, 2, 4, 6, 7, 8, 9, 11, 12, 24, 31, 32, 34, 40, 50, 52, 60, 62, 64, 68, 70, 73, 77, 78, 81, 85, 88, 89, 90, 94}
    ]
}

# === 辅助函数（全部整数） ===
def get_bounds():
    mins = np.array([EXPERIMENT_CONFIG['MIN_X'], EXPERIMENT_CONFIG['MIN_Y'], EXPERIMENT_CONFIG['MIN_Z']], dtype=np.float32)
    maxs = np.array([EXPERIMENT_CONFIG['MAX_X'], EXPERIMENT_CONFIG['MAX_Y'], EXPERIMENT_CONFIG['MAX_Z']], dtype=np.float32)
    return mins, maxs

def clip_state(state):
    mins, maxs = get_bounds()
    return np.clip(state, mins, maxs)

def denormalize_state(normalized_state):
    """将归一化状态还原为原始值（全部整数）"""
    mins, maxs = get_bounds()
    state = normalized_state * (maxs - mins) + mins
    # 全部四舍五入取整
    state = np.round(state).astype(int)
    # 确保在范围内
    state = np.clip(state, mins, maxs)
    return state

def coverage_similarity(triggered, target_path):
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0
    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)

def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']
    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']
    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']
    return reward

# ==================== 执行编排规则函数 ====================
def execute_orchestration_rules(a):
    path_depth, file_count, access_level = float(a[0]), int(a[1]), float(a[2])
    triggered = set()
    b = {}

    # 以下为原始编排规则（与用户提供一致）
    if (path_depth >= 10) != (path_depth >= 12):
        b[0] = 1; triggered.add(1)
    if (path_depth >= 10) != (path_depth == 10):
        b[1] = 2; triggered.add(2)
    if (path_depth >= 10) != (path_depth >= 8):
        b[2] = 3; triggered.add(3)

    if path_depth >= 10:
        if (file_count >= 10000 and access_level >= 3) != (file_count >= 10000 and access_level >= 4.3):
            b[3] = 4; triggered.add(4)
        if (file_count >= 10000 and access_level >= 3) != (file_count >= 10000 and access_level == 3):
            b[4] = 5; triggered.add(5)
        if (file_count >= 10000 and access_level >= 3) != (file_count >= 10000 and access_level >= 3.2):
            b[5] = 6; triggered.add(6)
        if (file_count >= 10000 and access_level >= 3) != (file_count == 10000 and access_level >= 3):
            b[6] = 7; triggered.add(7)
        if (file_count >= 10000 and access_level >= 3) != (file_count >= 10000 and access_level >= 4):
            b[7] = 8; triggered.add(8)
        if (file_count >= 10000 and access_level >= 3) != (file_count >= 13400 and access_level >= 3.9):
            b[8] = 9; triggered.add(9)
        if (file_count >= 10000 and access_level >= 3) != (file_count != 10000 and access_level >= 3):
            b[9] = 10; triggered.add(10)
        if (file_count >= 10000 and access_level >= 3) != (file_count >= 10000 and access_level >= 5):
            b[10] = 11; triggered.add(11)

        if file_count >= 10000 and access_level >= 3:
            if (access_level == 3 and file_count >= 50000) != (access_level != 3 and file_count >= 50000):
                b[11] = 12; triggered.add(12)
            if (access_level == 3 and file_count >= 50000) != (access_level >= 3 and file_count >= 50000):
                b[12] = 13; triggered.add(13)
            if (access_level == 3 and file_count >= 50000) != (access_level <= 3 and file_count >= 50000):
                b[13] = 14; triggered.add(14)
            if (access_level == 3 and file_count >= 50000) != (access_level == 3 or file_count >= 50000):
                b[14] = 15; triggered.add(15)

            if (file_count >= 25000) != (file_count != 25000):
                b[15] = 16; triggered.add(16)
            if (file_count >= 25000) != (file_count >= 30000):
                b[16] = 17; triggered.add(17)
            if (file_count >= 25000) != (file_count >= 29000):
                b[17] = 18; triggered.add(18)

        if (file_count >= 5000 and access_level >= 2) != (file_count >= 5000 and access_level >= 2.5):
            b[18] = 19; triggered.add(19)
        if (file_count >= 5000 and access_level >= 2) != (file_count >= 5000 or access_level >= 3.5):
            b[19] = 20; triggered.add(20)
        if (file_count >= 5000 and access_level >= 2) != (file_count >= 5000 and access_level != 2):
            b[20] = 21; triggered.add(21)
        if (file_count >= 5000 and access_level >= 2) != (file_count != 5000 and access_level >= 2):
            b[21] = 22; triggered.add(22)
        if (file_count >= 5000 and access_level >= 2) != (file_count >= 5000 and access_level >= 3):
            b[22] = 23; triggered.add(23)
        if (file_count >= 5000 and access_level >= 2) != (file_count >= 5000 and access_level >= 3.3):
            b[23] = 24; triggered.add(24)

        if (file_count >= 1000) != (file_count >= 43000):
            b[24] = 25; triggered.add(25)
        if (file_count >= 1000) != (file_count >= 25000):
            b[25] = 26; triggered.add(26)
        if (file_count >= 1000) != (file_count >= 50000):
            b[26] = 27; triggered.add(27)

    if (path_depth >= 6) != (path_depth >= 7):
        b[27] = 28; triggered.add(28)
    if (path_depth >= 6) != (path_depth != 6):
        b[28] = 29; triggered.add(29)
    if (path_depth >= 6) != (path_depth >= 7):
        b[29] = 30; triggered.add(30)

    elif path_depth >= 6:
        if (file_count >= 20000 and access_level >= 3) != (file_count >= 20000 and access_level >= 4):
            b[30] = 31; triggered.add(31)
        if (file_count >= 20000 and access_level >= 3) != (file_count >= 20000 and access_level != 3):
            b[31] = 32; triggered.add(32)
        if (file_count >= 20000 and access_level >= 3) != (file_count >= 20000 and access_level == 3):
            b[32] = 33; triggered.add(33)
        if (file_count >= 20000 and access_level >= 3) != (file_count == 20000 and access_level >= 3):
            b[33] = 34; triggered.add(34)
        if (file_count >= 20000 and access_level >= 3) != (file_count != 20000 and access_level >= 3):
            b[34] = 35; triggered.add(35)

        if (file_count >= 8000 and access_level >= 2) != (file_count >= 54000 and access_level >= 2):
            b[35] = 36; triggered.add(36)
        if (file_count >= 8000 and access_level >= 2) != (file_count != 8000 and access_level >= 2):
            b[36] = 37; triggered.add(37)
        if (file_count >= 8000 and access_level >= 2) != (file_count >= 45000 and access_level >= 2):
            b[37] = 38; triggered.add(38)
        if (file_count >= 8000 and access_level >= 2) != (file_count >= 8000 and access_level != 2):
            b[38] = 39; triggered.add(39)
        if (file_count >= 8000 and access_level >= 2) != (file_count >= 8000 and access_level >= 3.3):
            b[39] = 40; triggered.add(40)
        if (file_count >= 8000 and access_level >= 2) != (file_count >= 8000 and access_level >= 2.5):
            b[40] = 41; triggered.add(41)
        if (file_count >= 8000 and access_level >= 2) != (file_count >= 80500 and access_level >= 2):
            b[41] = 42; triggered.add(42)

        elif file_count >= 8000 and access_level >= 2:
            if (access_level == 3) != (access_level <= 3):
                b[42] = 43; triggered.add(43)
            if (access_level == 3) != (access_level >= 3):
                b[43] = 44; triggered.add(44)

        if (file_count >= 2000) != (file_count != 2000):
            b[44] = 45; triggered.add(45)
        if (file_count >= 2000) != (file_count >= 62000):
            b[45] = 46; triggered.add(46)

    if (path_depth >= 3) != (path_depth != 3):
        b[46] = 47; triggered.add(47)
    if (path_depth >= 3) != (path_depth >= 3.5):
        b[47] = 48; triggered.add(48)

    elif path_depth >= 3:
        if (file_count >= 15000 and access_level >= 2) != (file_count != 15000 and access_level >= 2):
            b[48] = 49; triggered.add(49)
        if (file_count >= 15000 and access_level >= 2) != (file_count == 15000 and access_level >= 2):
            b[49] = 50; triggered.add(50)
        if (file_count >= 15000 and access_level >= 2) != (file_count >= 15000 and access_level != 2):
            b[50] = 51; triggered.add(51)
        if (file_count >= 15000 and access_level >= 2) != (file_count >= 15000 and access_level == 2):
            b[51] = 52; triggered.add(52)
        if (file_count >= 15000 and access_level >= 2) != (file_count >= 15000 and access_level >= 2.5):
            b[52] = 53; triggered.add(53)
        if (file_count >= 15000 and access_level >= 2) != (file_count >= 17500 and access_level >= 2):
            b[53] = 54; triggered.add(54)
        if (file_count >= 15000 and access_level >= 2) != (file_count >= 33330 and access_level >= 2):
            b[54] = 55; triggered.add(55)

        if file_count >= 15000 and access_level >= 2:
            if (access_level == 3) != (access_level >= 3):
                b[55] = 56; triggered.add(56)
            if (access_level == 3) != (access_level <= 3):
                b[56] = 57; triggered.add(57)

        if (file_count >= 5000) != (file_count != 5000):
            b[57] = 58; triggered.add(58)
        if (file_count >= 5000) != (file_count >= 55000):
            b[58] = 59; triggered.add(59)

    # 文件数量维度处理
    if (file_count >= 100000) != (file_count >= 90000):
        b[59] = 60; triggered.add(60)
    if (file_count >= 100000) != (file_count == 100000):
        b[60] = 61; triggered.add(61)

    if file_count >= 100000:
        if (path_depth >= 8 and access_level >= 3) != (path_depth >= 8 and access_level != 3):
            b[61] = 62; triggered.add(62)
        if (path_depth >= 8 and access_level >= 3) != (path_depth >= 8 and access_level == 3):
            b[62] = 63; triggered.add(63)
        if (path_depth >= 8 and access_level >= 3) != (path_depth == 8 and access_level >= 3):
            b[63] = 64; triggered.add(64)
        if (path_depth >= 8 and access_level >= 3) != (path_depth != 8 and access_level >= 3):
            b[64] = 65; triggered.add(65)
        if (path_depth >= 8 and access_level >= 3) != (path_depth >= 9 and access_level >= 3):
            b[65] = 66; triggered.add(66)
        if (path_depth >= 8 and access_level >= 3) != (path_depth >= 11 and access_level >= 3):
            b[66] = 67; triggered.add(67)
        if (path_depth >= 8 and access_level >= 3) != (path_depth >= 15 and access_level >= 3):
            b[67] = 68; triggered.add(68)

        if (path_depth >= 5 and access_level >= 2) != (path_depth != 5 and access_level >= 2):
            b[68] = 69; triggered.add(69)
        if (path_depth >= 5 and access_level >= 2) != (path_depth == 5 and access_level >= 2):
            b[69] = 70; triggered.add(70)
        if (path_depth >= 5 and access_level >= 2) != (path_depth >= 5 and access_level >= 2.9):
            b[70] = 71; triggered.add(71)
        if (path_depth >= 5 and access_level >= 2) != (path_depth >= 5 and access_level != 2):
            b[71] = 72; triggered.add(72)
        if (path_depth >= 5 and access_level >= 2) != (path_depth >= 5 and access_level == 2):
            b[72] = 73; triggered.add(73)
        if (path_depth >= 5 and access_level >= 2) != (path_depth >= 5 and access_level >= 2.5):
            b[73] = 74; triggered.add(74)
        if (path_depth >= 5 and access_level >= 2) != (path_depth >= 7 and access_level >= 2):
            b[74] = 75; triggered.add(75)

    if (file_count >= 50000) != (file_count != 50000):
        b[75] = 76; triggered.add(76)
    if (file_count >= 50000) != (file_count == 50000):
        b[76] = 77; triggered.add(77)

    elif file_count >= 50000:
        if (access_level >= 3) != (access_level != 3):
            b[77] = 78; triggered.add(78)
        if (access_level >= 3) != (access_level == 3):
            b[78] = 79; triggered.add(79)

    if (file_count >= 10000) != (file_count != 10000):
        b[79] = 80; triggered.add(80)
    if (file_count >= 10000) != (file_count == 10000):
        b[80] = 81; triggered.add(81)

    # 访问级别维度的扫描序列规划
    if (access_level == 3) != (access_level <= 3):
        b[81] = 82; triggered.add(82)
    if (access_level == 3) != (access_level >= 3):
        b[82] = 83; triggered.add(83)

    if access_level == 3:
        if (path_depth >= 8 and file_count >= 20000) != (path_depth != 8 and file_count >= 20000):
            b[83] = 84; triggered.add(84)
        if (path_depth >= 8 and file_count >= 20000) != (path_depth == 8 and file_count >= 20000):
            b[84] = 85; triggered.add(85)
        if (path_depth >= 8 and file_count >= 20000) != (path_depth >= 7 and file_count >= 20000):
            b[85] = 86; triggered.add(86)
        if (path_depth >= 8 and file_count >= 20000) != (path_depth >= 8 and file_count != 20000):
            b[86] = 87; triggered.add(87)
        if (path_depth >= 8 and file_count >= 20000) != (path_depth >= 8 and file_count == 20000):
            b[87] = 88; triggered.add(88)
        if (path_depth >= 8 and file_count >= 20000) != (path_depth >= 11 and file_count == 20000):
            b[88] = 89; triggered.add(89)
        if (path_depth >= 8 and file_count >= 20000) != (path_depth >= 8 and file_count == 25500):
            b[89] = 90; triggered.add(90)

        if (path_depth >= 5) != (path_depth != 5):
            b[90] = 91; triggered.add(91)
        if (path_depth >= 5) != (path_depth >= 3.5):
            b[91] = 92; triggered.add(92)
        if (path_depth >= 5) != (path_depth >= 6):
            b[92] = 93; triggered.add(93)

    if (access_level == 2) != (access_level >= 2):
        b[93] = 94; triggered.add(94)
    if (access_level == 2) != (access_level <= 2):
        b[94] = 95; triggered.add(95)

    return triggered

# ===== 将编排规则函数赋给 execute_Tr =====
execute_Tr = execute_orchestration_rules

# === DQN网络 ===
class DQNNetwork(nn.Module):
    def __init__(self, action_size=30):
        super(DQNNetwork, self).__init__()
        hidden_dim = EXPERIMENT_CONFIG['HIDDEN_DIM']
        self.fc1 = nn.Linear(3, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)
        self.output = nn.Linear(hidden_dim, action_size)

    def forward(self, x):
        x = torch.relu(self.fc1(x))
        x = torch.relu(self.fc2(x))
        x = torch.relu(self.fc3(x))
        return self.output(x)

# === 路径经验回放 ===
class PathReplayBuffer:
    def __init__(self, path_idx, capacity=20000):
        self.path_idx = path_idx
        self.capacity = capacity
        self.buffer = deque(maxlen=capacity)
        self.similarities = deque(maxlen=capacity)

    def push(self, state, action, reward, next_state, done, similarity):
        self.buffer.append((state, action, reward, next_state, done))
        self.similarities.append(similarity)

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return None
        batch = random.sample(self.buffer, batch_size)
        state, action, reward, next_state, done = map(np.stack, zip(*batch))
        return (
            torch.FloatTensor(state).to(device),
            torch.LongTensor(action).to(device),
            torch.FloatTensor(reward).to(device),
            torch.FloatTensor(next_state).to(device),
            torch.BoolTensor(done).to(device)
        )

    def get_top_k(self, k=20):
        if len(self.buffer) == 0:
            return []
        samples_with_sim = list(zip(self.buffer, self.similarities))
        samples_with_sim.sort(key=lambda x: x[1], reverse=True)
        top_k = samples_with_sim[:k]
        results = []
        for (state, _, _, _, _), similarity in top_k:
            original_state = denormalize_state(state)
            results.append({
                'state': original_state,
                'similarity': similarity,
                'triggered': execute_Tr(*original_state)  # 解包三个值
            })
        return results

    def __len__(self):
        return len(self.buffer)

# === DQN智能体 ===
class ImprovedDQNAgent:
    def __init__(self, num_paths, action_size=30):
        self.action_size = action_size
        self.num_paths = num_paths
        self.epsilon = 0.9
        self.epsilon_min = 0.1
        self.epsilon_decay = 0.995

        self.q_network = DQNNetwork(action_size).to(device)
        self.target_network = DQNNetwork(action_size).to(device)

        lr = EXPERIMENT_CONFIG['LEARNING_RATE']
        self.optimizer = optim.Adam(self.q_network.parameters(), lr=lr)

        capacity = EXPERIMENT_CONFIG['REPLAY_BUFFER_CAPACITY']
        self.replay_buffers = {}
        for path_idx in range(num_paths):
            self.replay_buffers[path_idx] = PathReplayBuffer(path_idx, capacity)

        self.replay_train_count = 0
        self.update_target_network()

    def discrete_to_action_delta(self, action_idx):
        delta_values = [5, 3, 2, 1, 0.5, -0.5, -1, -2, -3, -5]
        if action_idx >= 30:
            action_idx = action_idx % 30
        dim = action_idx // 10
        delta_idx = action_idx % 10
        delta = delta_values[delta_idx]
        action_delta = np.zeros(3)
        action_delta[dim] = delta
        return action_delta

    def get_action(self, state):
        mins, maxs = get_bounds()
        normalized_state = (state - mins) / (maxs - mins)  # 直接归一化到 [0,1]
        if random.random() < self.epsilon:
            action_idx = random.randint(0, self.action_size - 1)
        else:
            state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)
            with torch.no_grad():
                q_values = self.q_network(state_tensor)
                action_idx = q_values.argmax().item()
        action_delta = self.discrete_to_action_delta(action_idx)
        return action_delta, action_idx

    def store_experience(self, path_idx, state, action_idx, reward, next_state, done, similarity):
        mins, maxs = get_bounds()
        normalized_state = (state - mins) / (maxs - mins)
        normalized_next_state = (next_state - mins) / (maxs - mins)
        self.replay_buffers[path_idx].push(
            normalized_state, action_idx, reward,
            normalized_next_state, done, similarity
        )

    def replay_train(self, path_idx):
        batch_size = EXPERIMENT_CONFIG['REPLAY_BATCH_SIZE']
        batch = self.replay_buffers[path_idx].sample(batch_size)
        if batch is None:
            return
        states, actions, rewards, next_states, dones = batch
        current_q_values = self.q_network(states).gather(1, actions.unsqueeze(1))
        with torch.no_grad():
            next_q_values = self.target_network(next_states).max(1)[0]
            target_q_values = rewards + (0.99 * next_q_values * ~dones)
        loss = F.mse_loss(current_q_values.squeeze(), target_q_values)
        self.optimizer.zero_grad()
        loss.backward()
        torch.nn.utils.clip_grad_norm_(self.q_network.parameters(), 1.0)
        self.optimizer.step()
        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay
        self.replay_train_count += 1
        if self.replay_train_count % 2 == 0:
            self.update_target_network()

    def update_target_network(self):
        self.target_network.load_state_dict(self.q_network.state_dict())

    def get_all_top_k(self, k=20):
        results = {}
        for path_idx in range(self.num_paths):
            results[path_idx] = self.replay_buffers[path_idx].get_top_k(k)
        return results

    def get_buffer_stats(self):
        stats = {}
        for path_idx in range(self.num_paths):
            stats[path_idx] = len(self.replay_buffers[path_idx])
        return stats

# === 性能指标计算 ===
def calculate_run_performance(run_idx, dqn_results, training_time, total_steps, update_count, agent):
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)
    total_reward = 0
    average_reward = 0
    convergence = 0
    environment_adaptability = 0
    generalization_ability = 0
    computational_efficiency = 0
    policy_update_frequency = 0
    all_similarities = []
    total_samples = 0
    all_rewards = []
    for path_idx in range(num_paths):
        samples = dqn_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']
            total_reward += reward
            all_rewards.append(reward)
            all_similarities.append(similarity)
            total_samples += 1
    total_reward = total_reward
    if total_samples > 0:
        average_reward = total_reward / total_samples
    if all_similarities:
        convergence = np.mean(all_similarities)
    if len(all_similarities) > 1:
        environment_adaptability = 1 / (np.std(all_similarities) + 1e-8)
    generalization_ability = convergence
    if training_time > 0:
        computational_efficiency = total_steps / training_time
    if training_time > 0:
        policy_update_frequency = update_count / training_time
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0
    return {
        'Run': run_idx + 1,
        'Total Reward': round(total_reward, 2),
        'Average Reward': round(average_reward, 4),
        'Convergence': round(convergence, 4),
        'Environment Adaptability': round(environment_adaptability, 4),
        'Generalization Ability': round(generalization_ability, 4),
        'Computational Efficiency': round(computational_efficiency, 2),
        'Policy Update Frequency': round(policy_update_frequency, 4),
        'Average Similarity': round(avg_similarity, 4),
        'Max Similarity': round(max_similarity, 4),
        'Min Similarity': round(min_similarity, 4),
    }

# === 导出Excel ===
def export_to_excel(all_dqn_results, all_performance_data, target_paths, output_path="DQN_20_run.xlsx"):
    print("\n正在生成Excel...")
    all_dqn_summary_data = []
    all_dqn_detailed_data = []
    for run_idx, (dqn_results, performance_data) in enumerate(zip(all_dqn_results, all_performance_data)):
        dqn_summary_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = dqn_results[path_idx]
            if len(samples) == 0:
                dqn_summary_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Target Rule Count': len(target_path),
                    'Sample Count': 0,
                    'Average Similarity': 0,
                    'Max Similarity': 0,
                    'Min Similarity': 0,
                    'Similarity Std': 0,
                    'Perfect Coverage': 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path)))
                })
                continue
            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = 'Yes' if perfect_count > 0 else 'No'
            dqn_summary_data.append({
                'Run': run_idx + 1,
                'Path ID': path_idx + 1,
                'Target Rule Count': len(target_path),
                'Sample Count': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Max Similarity': round(max(similarities), 4),
                'Min Similarity': round(min(similarities), 4),
                'Similarity Std': round(np.std(similarities), 4),
                'Perfect Coverage': is_perfect,
                'Target Paths': ', '.join(map(str, sorted(target_path)))
            })
        all_dqn_summary_data.extend(dqn_summary_data)
        dqn_detailed_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = dqn_results[path_idx]
            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']
                dqn_detailed_data.append({
                    'Run': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Sample ID': sample_idx + 1,
                    'Path_Depth': int(state[0]),
                    'File_Count': int(state[1]),
                    'Access_Level': int(state[2]),
                    'Similarity': round(similarity, 4),
                    'Perfect Coverage': 'Yes' if abs(similarity - 1.0) < 0.001 else 'No',
                    'Target Paths': ', '.join(map(str, sorted(target_path))),
                    'Triggered Rules': ', '.join(map(str, sorted(triggered))),
                    'Hit Rule Count': len(target_path.intersection(triggered)),
                    'Target Rule Count': len(target_path)
                })
        all_dqn_detailed_data.extend(dqn_detailed_data)

    dqn_summary_df = pd.DataFrame(all_dqn_summary_data)
    dqn_detailed_df = pd.DataFrame(all_dqn_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        dqn_summary_df.to_excel(writer, sheet_name='DQN路径摘要', index=False)
        dqn_detailed_df.to_excel(writer, sheet_name='DQN详细样本数据', index=False)
        selected_columns = [
            'Run',
            'Total Reward', 'Average Reward', 'Convergence',
            'Environment Adaptability', 'Generalization Ability',
            'Computational Efficiency', 'Policy Update Frequency',
            'Average Similarity', 'Max Similarity', 'Min Similarity'
        ]
        performance_df_selected = performance_df[selected_columns]
        performance_df_selected.to_excel(writer, sheet_name='性能指标', index=False)
        workbook = writer.book
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        perfect_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        ws1 = writer.sheets['DQN路径摘要']
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row_idx in range(2, ws1.max_row + 1):
            if ws1.cell(row_idx, 9).value == 'Yes':
                for col_idx in range(1, ws1.max_column + 1):
                    ws1.cell(row_idx, col_idx).fill = perfect_fill
        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 15
        ws1.column_dimensions['F'].width = 15
        ws1.column_dimensions['G'].width = 15
        ws1.column_dimensions['H'].width = 15
        ws1.column_dimensions['I'].width = 15
        ws1.column_dimensions['J'].width = 50

        ws2 = writer.sheets['DQN详细样本数据']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 14
        ws2.column_dimensions['E'].width = 14
        ws2.column_dimensions['F'].width = 14
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 15
        ws2.column_dimensions['I'].width = 40
        ws2.column_dimensions['J'].width = 40
        ws2.column_dimensions['K'].width = 15
        ws2.column_dimensions['L'].width = 15

        ws3 = writer.sheets['性能指标']
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        for col in columns:
            ws3.column_dimensions[col].width = 18

    print(f"Excel已保存: {output_path}")
    print(f"  - Sheet1: DQN路径摘要 ({len(all_dqn_summary_data)} 行)")
    print(f"  - Sheet2: DQN详细样本数据 ({len(all_dqn_detailed_data)} 行)")
    print(f"  - Sheet3: 性能指标 ({len(all_performance_data)} 行)")

# === 训练流程 ===
def train_dqn_workflow():
    print("=" * 80)
    print("DQN训练开始 (基于覆盖相似度 - 编排规则)")
    print(f"每路径训练轮数: {EXPERIMENT_CONFIG['NUM_ROUNDS']}, 每样本步数: {EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(f"经验回放容量: {EXPERIMENT_CONFIG['REPLAY_BUFFER_CAPACITY']}")
    print("=" * 80)

    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)
    agent = ImprovedDQNAgent(num_paths=num_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n生成初始样本: 每路径 {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']} 个")
    path_samples = {}
    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            state = np.array([
                random.randint(EXPERIMENT_CONFIG['MIN_X'], EXPERIMENT_CONFIG['MAX_X']),
                random.randint(EXPERIMENT_CONFIG['MIN_Y'], EXPERIMENT_CONFIG['MAX_Y']),
                random.randint(EXPERIMENT_CONFIG['MIN_Z'], EXPERIMENT_CONFIG['MAX_Z'])
            ], dtype=np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  路径 {path_idx + 1}/{num_paths}: 已生成 {len(samples)} 个样本")

    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size
    num_rounds = EXPERIMENT_CONFIG['NUM_ROUNDS']

    print(f"\n训练配置:")
    print(f"  - 每批样本数: {batch_size}")
    print(f"  - 每路径批次数: {num_batches}")
    print(f"  - 每路径轮数: {num_rounds}")
    print(f"  - 每样本步数: {EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(f"  - 总训练批次: {num_paths * num_rounds * num_batches} 批")
    print("-" * 80)

    for path_idx in range(num_paths):
        target_path = target_paths[path_idx]
        print(f"\n{'=' * 80}")
        print(f"开始训练路径 {path_idx + 1}/{num_paths}")
        print(f"目标规则集: {sorted(target_path)}")
        print(f"{'=' * 80}")

        for round_idx in range(num_rounds):
            print(f"\n  路径 {path_idx + 1} - 第 {round_idx + 1}/{num_rounds} 轮")
            for batch_idx in range(num_batches):
                batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]
                batch_rewards = []
                batch_similarities = []

                for sample_idx, initial_state in enumerate(batch_samples):
                    state = initial_state.copy()
                    episode_reward = 0
                    final_similarity = 0

                    for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                        action_delta, action_idx = agent.get_action(state)
                        next_state = state + action_delta
                        next_state = clip_state(next_state)
                        triggered = execute_Tr(*next_state)
                        reward = unified_reward_function(triggered, target_path)
                        similarity = coverage_similarity(triggered, target_path)
                        done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)
                        agent.store_experience(
                            path_idx, state, action_idx, reward, next_state, done, similarity
                        )
                        state = next_state
                        episode_reward += reward
                        final_similarity = similarity
                        total_steps += 1

                    batch_rewards.append(episode_reward)
                    batch_similarities.append(final_similarity)

                avg_reward = np.mean(batch_rewards)
                avg_similarity = np.mean(batch_similarities)
                max_similarity = np.max(batch_similarities)
                print(f"    批次 {batch_idx + 1}/{num_batches}: 平均奖励={avg_reward:.2f}, 平均相似度={avg_similarity:.4f}, 最大相似度={max_similarity:.4f}, epsilon={agent.epsilon:.3f}")

                print(f"    进行路径 {path_idx + 1} 的经验回放训练...")
                agent.replay_train(path_idx)

                buffer_size = len(agent.replay_buffers[path_idx])
                print(f"    路径 {path_idx + 1} 经验池大小: {buffer_size}, 总回放次数: {agent.replay_train_count}")

    training_time = time.time() - start_time

    print("\n" + "=" * 80)
    print(f"DQN训练完成! 总耗时: {training_time:.2f} 秒, 总步数: {total_steps}")
    print(f"总回放训练次数: {agent.replay_train_count}")
    print(f"目标网络更新次数: {agent.replay_train_count // 2}")

    print("\n路径经验池状态:")
    buffer_stats = agent.get_buffer_stats()
    for path_idx, size in buffer_stats.items():
        print(f"  路径 {path_idx + 1}: {size} 条经验")

    print("=" * 80)

    print(f"\n提取每路径 Top-{EXPERIMENT_CONFIG['TOP_K_SAMPLES']} 高相似度样本...")
    dqn_top_k_results = agent.get_all_top_k(EXPERIMENT_CONFIG['TOP_K_SAMPLES'])

    return agent, dqn_top_k_results, training_time, total_steps, agent.replay_train_count

# === 主函数 ===
def main():
    print("\n" + "=" * 80)
    print("DQN - 20次运行实验 (编排规则)")
    print("性能指标：覆盖相似度")
    print(f"状态范围: path_depth[{EXPERIMENT_CONFIG['MIN_X']}-{EXPERIMENT_CONFIG['MAX_X']}], "
          f"file_count[{EXPERIMENT_CONFIG['MIN_Y']}-{EXPERIMENT_CONFIG['MAX_Y']}], "
          f"access_level[{EXPERIMENT_CONFIG['MIN_Z']}-{EXPERIMENT_CONFIG['MAX_Z']}]")
    print("=" * 80)

    all_dqn_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'='*80}")
        print(f"开始第 {run_idx + 1}/{EXPERIMENT_CONFIG['NUM_RUNS']} 次运行")
        print(f"{'='*80}")

        dqn_agent, dqn_results, training_time, total_steps, update_count = train_dqn_workflow()
        performance_data = calculate_run_performance(
            run_idx, dqn_results, training_time, total_steps, update_count, dqn_agent
        )

        all_dqn_results.append(dqn_results)
        all_performance_data.append(performance_data)

        print(f"\n第 {run_idx + 1} 次运行完成!")
        print(f"  总奖励: {performance_data['Total Reward']}")
        print(f"  平均奖励: {performance_data['Average Reward']}")
        print(f"  收敛度: {performance_data['Convergence']}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"DQN_20_runs_orchestration_{timestamp}.xlsx"
    export_to_excel(all_dqn_results, all_performance_data, target_paths, output_path)

    print("\n" + "=" * 80)
    print("20次运行统计汇总")
    print("=" * 80)
    total_rewards = [p['Total Reward'] for p in all_performance_data]
    average_rewards = [p['Average Reward'] for p in all_performance_data]
    convergences = [p['Convergence'] for p in all_performance_data]
    environment_adaptabilities = [p['Environment Adaptability'] for p in all_performance_data]
    generalization_abilities = [p['Generalization Ability'] for p in all_performance_data]
    computational_efficiencies = [p['Computational Efficiency'] for p in all_performance_data]
    policy_update_frequencies = [p['Policy Update Frequency'] for p in all_performance_data]
    avg_similarities = [p['Average Similarity'] for p in all_performance_data]

    print(f"\n总奖励:")
    print(f"  平均值: {np.mean(total_rewards):.2f}")
    print(f"  标准差: {np.std(total_rewards):.2f}")

    print(f"\n平均奖励:")
    print(f"  平均值: {np.mean(average_rewards):.4f}")
    print(f"  标准差: {np.std(average_rewards):.4f}")

    print(f"\n收敛度 (平均相似度):")
    print(f"  平均值: {np.mean(convergences):.4f}")
    print(f"  标准差: {np.std(convergences):.4f}")

    print(f"\n环境适应性:")
    print(f"  平均值: {np.mean(environment_adaptabilities):.4f}")
    print(f"  标准差: {np.std(environment_adaptabilities):.4f}")

    print(f"\n泛化能力:")
    print(f"  平均值: {np.mean(generalization_abilities):.4f}")
    print(f"  标准差: {np.std(generalization_abilities):.4f}")

    print(f"\n计算效率 (步/秒):")
    print(f"  平均值: {np.mean(computational_efficiencies):.2f}")
    print(f"  标准差: {np.std(computational_efficiencies):.2f}")

    print(f"\n策略更新频率 (次/秒):")
    print(f"  平均值: {np.mean(policy_update_frequencies):.4f}")
    print(f"  标准差: {np.std(policy_update_frequencies):.4f}")

    print(f"\n平均相似度统计:")
    print(f"  平均值: {np.mean(avg_similarities):.4f}")
    print(f"  标准差: {np.std(avg_similarities):.4f}")

    print("\n" + "=" * 80)
    print(f"所有 {EXPERIMENT_CONFIG['NUM_RUNS']} 次运行完成!")
    print("=" * 80)

if __name__ == "__main__":
    main()