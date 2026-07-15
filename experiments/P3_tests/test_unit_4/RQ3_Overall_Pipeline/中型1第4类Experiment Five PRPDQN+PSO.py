import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

NUM_RUNS = 20  # 实验运行次数

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === 第四个单元测试的三维范围（威胁分析） ===
THREAT_COUNT_MIN = 1
THREAT_COUNT_MAX = 120
SEVERITY_LEVEL_MIN = 1.0
SEVERITY_LEVEL_MAX = 10.0
CONFIDENCE_SCORE_MIN = 1.0
CONFIDENCE_SCORE_MAX = 100.0

BOUNDS = {
    'threat_count': (THREAT_COUNT_MIN, THREAT_COUNT_MAX),
    'severity_level': (SEVERITY_LEVEL_MIN, SEVERITY_LEVEL_MAX),
    'confidence_score': (CONFIDENCE_SCORE_MIN, CONFIDENCE_SCORE_MAX)
}

# === 动作增量（沿用原有数值，执行时自动裁剪） ===
DELTA_THREAT = [35, 25, 10, 5, 2, -2, -5, -10, -25, -35]
DELTA_SEVERITY = [35, 25, 10, 5, 2, -2, -5, -10, -25, -35]
DELTA_CONFIDENCE = [35, 25, 10, 5, 2, -2, -5, -10, -25, -35]   # 实际会被裁剪到边界内

# ========================================
# ========== 状态归一化函数 ==========
# ========================================
def normalize_state(state):
    """
    将状态归一化到 [0,1]
    state: (threat_count, severity_level, confidence_score)
    """
    normalized = np.array([
        (state[0] - BOUNDS['threat_count'][0]) / (BOUNDS['threat_count'][1] - BOUNDS['threat_count'][0]),
        (state[1] - BOUNDS['severity_level'][0]) / (BOUNDS['severity_level'][1] - BOUNDS['severity_level'][0]),
        (state[2] - BOUNDS['confidence_score'][0]) / (BOUNDS['confidence_score'][1] - BOUNDS['confidence_score'][0])
    ], dtype=np.float32)
    return normalized

def denormalize_state(normalized_state):
    """
    从归一化状态还原为原始状态
    """
    state = np.array([
        normalized_state[0] * (BOUNDS['threat_count'][1] - BOUNDS['threat_count'][0]) + BOUNDS['threat_count'][0],
        normalized_state[1] * (BOUNDS['severity_level'][1] - BOUNDS['severity_level'][0]) + BOUNDS['severity_level'][0],
        normalized_state[2] * (BOUNDS['confidence_score'][1] - BOUNDS['confidence_score'][0]) + BOUNDS['confidence_score'][0]
    ])
    return state

# ========================================

def generate_random_state():
    """生成随机状态（整数或浮点数）"""
    count = np.random.randint(BOUNDS['threat_count'][0], BOUNDS['threat_count'][1] + 1)
    severity = np.random.uniform(BOUNDS['severity_level'][0], BOUNDS['severity_level'][1])
    confidence = np.random.uniform(BOUNDS['confidence_score'][0], BOUNDS['confidence_score'][1])
    return np.array([count, severity, confidence])

def clip_state(state):
    """将状态裁剪到边界内"""
    return np.array([
        np.clip(state[0], BOUNDS['threat_count'][0], BOUNDS['threat_count'][1]),
        np.clip(state[1], BOUNDS['severity_level'][0], BOUNDS['severity_level'][1]),
        np.clip(state[2], BOUNDS['confidence_score'][0], BOUNDS['confidence_score'][1])
    ])

def is_state_valid(state):
    """检查状态是否在边界内"""
    return (BOUNDS['threat_count'][0] <= state[0] <= BOUNDS['threat_count'][1] and
            BOUNDS['severity_level'][0] <= state[1] <= BOUNDS['severity_level'][1] and
            BOUNDS['confidence_score'][0] <= state[2] <= BOUNDS['confidence_score'][1])

# ========================================
# ========== 执行威胁分析规则函数 ==========
def execute_threat_analysis_rules(a):
    """
    威胁分析规则执行函数
    参数a: (threat_count, severity_level, confidence_score)
    返回: 触发的规则编号集合
    """
    threat_count, severity_level, confidence_score = int(a[0]), float(a[1]), float(a[2])
    triggered = set()
    b = {}

    # 按照原始代码的逻辑进行条件判断
    if (threat_count >= 100) != (threat_count >= 110):
        b[0] = 1
        triggered.add(1)
    if (threat_count >= 100) != (threat_count == 100):
        b[1] = 2
        triggered.add(2)
    if (threat_count >= 100) != (threat_count >= 75):
        b[2] = 3
        triggered.add(3)

    if threat_count >= 100:
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 8.0 and confidence_score != 90):
            b[3] = 4
            triggered.add(4)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 8.0 and confidence_score >= 95):
            b[4] = 5
            triggered.add(5)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level != 8.0 and confidence_score >= 90):
            b[5] = 6
            triggered.add(6)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 9.5 and confidence_score >= 90):
            b[6] = 7
            triggered.add(7)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 9.3 and confidence_score >= 90):
            b[7] = 8
            triggered.add(8)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 8.0 and confidence_score >= 88):
            b[8] = 9
            triggered.add(9)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 8.0 and confidence_score >= 95):
            b[9] = 10
            triggered.add(10)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 8.3 and confidence_score >= 90):
            b[10] = 11
            triggered.add(11)

        if severity_level >= 8.0 and confidence_score >= 90:
            if (confidence_score >= 95) != (confidence_score >= 93):
                b[11] = 12
                triggered.add(12)
            if (confidence_score >= 95) != (confidence_score == 95):
                b[12] = 13
                triggered.add(13)
            if (confidence_score >= 95) != (confidence_score >= 97):
                b[13] = 14
                triggered.add(14)

            if (severity_level >= 9.0) != (severity_level >= 9.1):
                b[14] = 15
                triggered.add(15)
            if (severity_level >= 9.0) != (severity_level == 9.0):
                b[15] = 16
                triggered.add(16)
            if (severity_level >= 9.0) != (severity_level >= 8.0):
                b[16] = 17
                triggered.add(17)

        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level != 6.0 and confidence_score >= 80):
            b[17] = 18
            triggered.add(18)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level == 6.0 and confidence_score >= 80):
            b[18] = 19
            triggered.add(19)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score != 80):
            b[19] = 20
            triggered.add(20)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score == 80):
            b[20] = 21
            triggered.add(21)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score >= 82):
            b[21] = 22
            triggered.add(22)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score >= 88):
            b[22] = 23
            triggered.add(23)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 7.7 and confidence_score >= 82):
            b[23] = 24
            triggered.add(24)

        if (severity_level >= 4.0) != (severity_level != 4.0):
            b[24] = 25
            triggered.add(25)
        if (severity_level >= 4.0) != (severity_level >= 7.0):
            b[25] = 26
            triggered.add(26)
        if (severity_level >= 4.0) != (severity_level >= 5.0):
            b[26] = 27
            triggered.add(27)

    if (threat_count >= 50) != (threat_count >= 75):
        b[27] = 28
        triggered.add(28)
    if (threat_count >= 50) != (threat_count >= 59):
        b[28] = 29
        triggered.add(29)
    if (threat_count >= 50) != (threat_count >= 55):
        b[29] = 30
        triggered.add(30)

    elif threat_count >= 50:
        if (severity_level >= 7.0 and confidence_score >= 85) != (severity_level != 7.0 and confidence_score >= 85):
            b[30] = 31
            triggered.add(31)
        if (severity_level >= 7.0 and confidence_score >= 85) != (severity_level == 7.0 and confidence_score >= 85):
            b[31] = 32
            triggered.add(32)
        if (severity_level >= 7.0 and confidence_score >= 85) != (severity_level >= 7.0 and confidence_score != 85):
            b[32] = 33
            triggered.add(33)
        if (severity_level >= 7.0 and confidence_score >= 85) != (severity_level >= 7.0 and confidence_score == 85):
            b[33] = 34
            triggered.add(34)
        if (severity_level >= 7.0 and confidence_score >= 85) != (severity_level >= 7.0 and confidence_score >= 88):
            b[34] = 35
            triggered.add(35)

        if severity_level >= 7.0 and confidence_score >= 85:
            if (confidence_score >= 95) != (confidence_score >= 85):
                b[35] = 36
                triggered.add(36)
            if (confidence_score >= 95) != (confidence_score == 95):
                b[36] = 37
                triggered.add(37)

        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level >= 5.0 and confidence_score != 70):
            b[37] = 38
            triggered.add(38)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level >= 5.0 and confidence_score == 70):
            b[38] = 39
            triggered.add(39)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level >= 5.0 and confidence_score >= 75):
            b[39] = 40
            triggered.add(40)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level != 5.0 and confidence_score >= 70):
            b[40] = 41
            triggered.add(41)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level == 5.0 and confidence_score >= 70):
            b[41] = 42
            triggered.add(42)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level >= 5.0 and confidence_score >= 74):
            b[42] = 43
            triggered.add(43)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level >= 6.2 and confidence_score >= 70):
            b[43] = 44
            triggered.add(44)

    if (threat_count >= 20) != (threat_count != 20):
        b[44] = 45
        triggered.add(45)
    if (threat_count >= 20) != (threat_count >= 30):
        b[45] = 46
        triggered.add(46)
    if (threat_count >= 20) != (threat_count >= 25):
        b[46] = 47
        triggered.add(47)

    elif threat_count >= 20:
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score != 80):
            b[47] = 48
            triggered.add(48)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score == 80):
            b[48] = 49
            triggered.add(49)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score >= 87):
            b[49] = 50
            triggered.add(50)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level != 6.0 and confidence_score >= 80):
            b[50] = 51
            triggered.add(51)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level == 6.0 and confidence_score >= 80):
            b[51] = 52
            triggered.add(52)

        if (severity_level >= 4.0) != (severity_level != 4.0):
            b[52] = 53
            triggered.add(53)
        if (severity_level >= 4.0) != (severity_level >= 2.5):
            b[53] = 54
            triggered.add(54)

    if (threat_count >= 5) != (threat_count != 5):
        b[54] = 55
        triggered.add(55)
    if (threat_count >= 5) != (threat_count >= 7.2):
        b[55] = 56
        triggered.add(56)

    elif threat_count >= 5:
        if (severity_level >= 7.0) != (severity_level >= 8.2):
            b[56] = 57
            triggered.add(57)
        if (severity_level >= 7.0) != (severity_level == 7.0):
            b[57] = 58
            triggered.add(58)
        if (severity_level >= 7.0) != (severity_level >= 7.5):
            b[58] = 59
            triggered.add(59)

    # 严重程度维度的利益相关者通知
    if (severity_level >= 9.0) != (severity_level >= 9.4):
        b[59] = 60
        triggered.add(60)
    if (severity_level >= 9.0) != (severity_level == 9.0):
        b[60] = 61
        triggered.add(61)

    if severity_level >= 9.0:
        if (threat_count >= 50 and confidence_score >= 90) != (threat_count >= 50 and confidence_score >= 93):
            b[61] = 62
            triggered.add(62)
        if (threat_count >= 50 and confidence_score >= 90) != (threat_count >= 50 and confidence_score == 90):
            b[62] = 63
            triggered.add(63)
        if (threat_count >= 50 and confidence_score >= 90) != (threat_count >= 55 and confidence_score >= 93):
            b[63] = 64
            triggered.add(64)
        if (threat_count >= 50 and confidence_score >= 90) != (threat_count >= 66 and confidence_score >= 90):
            b[64] = 65
            triggered.add(65)
        if (threat_count >= 50 and confidence_score >= 90) != (threat_count >= 50 and confidence_score == 90):
            b[65] = 66
            triggered.add(66)

        if (threat_count >= 20 or confidence_score >= 85) != (threat_count != 20 or confidence_score >= 85):
            b[66] = 67
            triggered.add(67)
        if (threat_count >= 20 or confidence_score >= 85) != (threat_count >= 50 or confidence_score >= 85):
            b[67] = 68
            triggered.add(68)
        if (threat_count >= 20 or confidence_score >= 85) != (threat_count >= 35 or confidence_score >= 85):
            b[68] = 69
            triggered.add(69)
        if (threat_count >= 20 or confidence_score >= 85) != (threat_count >= 20 or confidence_score != 85):
            b[69] = 70
            triggered.add(70)
        if (threat_count >= 20 or confidence_score >= 85) != (threat_count >= 20 or confidence_score == 85):
            b[70] = 71
            triggered.add(71)

    if (severity_level >= 7.0) != (severity_level >= 5.5):
        b[71] = 72
        triggered.add(72)
    if (severity_level >= 7.0) != (severity_level == 7.0):
        b[72] = 73
        triggered.add(73)

    elif severity_level >= 7.0:
        if (confidence_score >= 85) != (confidence_score >= 87):
            b[73] = 74
            triggered.add(74)
        if (confidence_score >= 85) != (confidence_score == 85):
            b[74] = 75
            triggered.add(75)

    if (severity_level >= 5.0) != (severity_level != 5.0):
        b[75] = 76
        triggered.add(76)
    if (severity_level >= 5.0) != (severity_level >= 6.0):
        b[76] = 77
        triggered.add(77)

    # 置信度分数维度的行动建议
    if (confidence_score >= 95) != (confidence_score >= 97):
        b[77] = 78
        triggered.add(78)
    if (confidence_score >= 95) != (confidence_score >= 89):
        b[78] = 79
        triggered.add(79)
    if (confidence_score >= 95) != (confidence_score == 95):
        b[79] = 80
        triggered.add(80)

    if confidence_score >= 95:
        if (severity_level >= 8.0 and threat_count >= 30) != (severity_level >= 8.0 and threat_count != 30):
            b[80] = 81
            triggered.add(81)
        if (severity_level >= 8.0 and threat_count >= 30) != (severity_level >= 8.0 and threat_count == 30):
            b[81] = 82
            triggered.add(82)
        if (severity_level >= 8.0 and threat_count >= 30) != (severity_level >= 8.0 and threat_count >= 31):
            b[82] = 83
            triggered.add(83)
        if (severity_level >= 8.0 and threat_count >= 30) != (severity_level >= 8.0 and threat_count >= 45):
            b[83] = 84
            triggered.add(84)
        if (severity_level >= 8.0 and threat_count >= 30) != (severity_level == 8.0 and threat_count >= 30):
            b[84] = 85
            triggered.add(85)

        if (severity_level >= 6.0) != (severity_level >= 6.6):
            b[85] = 86
            triggered.add(86)
        if (severity_level >= 6.0) != (severity_level >= 7.3):
            b[86] = 87
            triggered.add(87)

    if (confidence_score >= 85) != (confidence_score >= 78):
        b[87] = 88
        triggered.add(88)
    if (confidence_score >= 85) != (confidence_score >= 87):
        b[88] = 89
        triggered.add(89)

    elif confidence_score >= 85:
        if (severity_level >= 7.0) != (severity_level >= 8.7):
            b[89] = 90
            triggered.add(90)
        if (severity_level >= 7.0) != (severity_level == 7.0):
            b[90] = 91
            triggered.add(91)
        if (severity_level >= 7.0) != (severity_level >= 7.2):
            b[91] = 92
            triggered.add(92)

        if (severity_level >= 5.0) != (severity_level >= 6.3):
            b[92] = 93
            triggered.add(93)
        if (severity_level >= 5.0) != (severity_level >= 3.7):
            b[93] = 94
            triggered.add(94)

    if (confidence_score >= 70) != (confidence_score >= 77):
        b[94] = 95
        triggered.add(95)
    if (confidence_score >= 70) != (confidence_score >= 86):
        b[95] = 96
        triggered.add(96)

    return triggered

# 将执行函数赋给 execute_Tr（供后续代码调用）
execute_Tr = execute_threat_analysis_rules

# === 目标路径组（威胁分析） ===
targetPaths = [
    # A1
    {6, 7, 8, 11, 13, 17, 19, 21, 28, 29, 30, 32, 34, 37, 39, 42, 49, 52, 57, 58, 63, 64, 65, 66, 73, 75, 80, 82, 90, 91},
    # A2
    {5, 7, 8, 10, 15, 19, 21, 28, 29, 32, 34, 36, 39, 42, 49, 52, 58, 60, 62, 63, 64, 65, 66, 73, 75, 79, 82, 85, 91},
    # A3
    {6, 7, 8, 11, 14, 17, 19, 21, 28, 29, 30, 32, 34, 39, 42, 49, 52, 57, 58, 63, 64, 65, 66, 73, 75, 78, 82, 90, 91},
    # A4
    {4, 5, 7, 8, 10, 15, 19, 21, 28, 29, 30, 32, 34, 36, 39, 42, 49, 52, 58, 60, 62, 64, 65, 73, 75, 79, 82, 85, 91},
    # A5
    {5, 6, 7, 8, 10, 11, 12, 17, 19, 21, 28, 32, 34, 36, 39, 42, 49, 52, 57, 58, 63, 65, 66, 73, 75, 79, 82, 90, 91},
    # A6
    {5, 10, 16, 19, 21, 28, 29, 30, 32, 34, 36, 39, 42, 49, 52, 58, 61, 62, 63, 64, 65, 66, 73, 75, 79, 82, 85, 91},
    # A7
    {1, 2, 6, 7, 8, 11, 13, 14, 17, 19, 21, 32, 34, 37, 39, 42, 49, 52, 57, 58, 63, 66, 73, 75, 78, 80, 82, 90, 91},
    # A8
    {5, 6, 7, 8, 10, 11, 12, 17, 19, 21, 32, 34, 36, 39, 42, 45, 49, 52, 55, 57, 58, 71, 73, 75, 79, 81, 90, 91},
    # A9
    {3, 6, 7, 8, 11, 13, 14, 17, 19, 21, 32, 34, 37, 39, 42, 49, 52, 57, 58, 63, 66, 73, 75, 78, 80, 82, 90, 91},
    # A10
    {4, 5, 6, 7, 8, 10, 11, 17, 19, 21, 32, 34, 36, 39, 42, 45, 46, 47, 49, 52, 57, 58, 73, 75, 79, 81, 90, 91},
    # A11
    {6, 7, 8, 11, 13, 17, 19, 21, 32, 34, 37, 39, 42, 45, 49, 52, 55, 56, 57, 58, 71, 73, 75, 80, 81, 90, 91},
    # A12
    {4, 15, 19, 21, 23, 28, 29, 30, 32, 34, 35, 36, 39, 42, 49, 50, 52, 58, 60, 73, 74, 75, 82, 85, 89, 91},
    # A13
    {4, 17, 19, 21, 23, 32, 33, 35, 36, 39, 42, 45, 46, 47, 49, 50, 52, 57, 58, 73, 74, 81, 89, 90, 91, 96},
    # A14
    {4, 17, 19, 21, 23, 32, 33, 35, 36, 39, 42, 45, 49, 50, 52, 56, 57, 58, 70, 73, 74, 81, 89, 90, 91, 96},
    # A15
    {4, 17, 19, 21, 23, 33, 39, 42, 45, 46, 47, 49, 50, 52, 57, 58, 67, 68, 69, 73, 81, 88, 90, 91, 96},
    # A16
    {4, 17, 19, 21, 22, 23, 24, 33, 39, 42, 45, 49, 50, 52, 55, 57, 58, 67, 70, 73, 81, 88, 90, 91, 96},
    # A17
    {6, 13, 14, 18, 21, 24, 26, 28, 31, 37, 39, 42, 44, 49, 51, 63, 65, 66, 72, 75, 78, 80, 86, 87, 93},
    # A18
    {6, 13, 14, 19, 21, 24, 28, 31, 34, 37, 39, 42, 49, 52, 57, 59, 63, 65, 66, 75, 78, 80, 87, 90, 92},
    # A19
    {5, 6, 7, 8, 10, 11, 17, 19, 21, 32, 34, 36, 39, 42, 49, 52, 57, 58, 73, 75, 79, 82, 84, 90, 91},
    # A20
    {4, 16, 19, 20, 22, 23, 24, 28, 29, 30, 33, 39, 42, 48, 50, 52, 58, 61, 73, 82, 85, 88, 91, 96},
    # A21
    {6, 13, 14, 18, 26, 28, 29, 30, 31, 37, 39, 41, 44, 51, 63, 64, 65, 66, 75, 76, 77, 78, 80, 93},
    # A22
    {6, 7, 8, 11, 14, 17, 19, 21, 32, 34, 39, 42, 49, 52, 57, 58, 73, 75, 78, 81, 83, 84, 90, 91},
    # A23
    {4, 9, 17, 19, 21, 32, 34, 36, 39, 42, 45, 46, 47, 49, 52, 57, 58, 73, 75, 79, 81, 90, 91},
    # A24
    {4, 17, 20, 33, 38, 40, 42, 43, 45, 48, 55, 56, 57, 58, 67, 70, 73, 81, 90, 91, 95, 96},
    # A25
    {6, 18, 25, 26, 27, 28, 29, 30, 31, 36, 41, 51, 53, 62, 63, 64, 65, 66, 75, 76, 79, 94},
    # A26
    {6, 12, 18, 25, 28, 29, 30, 31, 36, 41, 51, 53, 54, 63, 64, 65, 66, 75, 76, 79}
]

def jaccard_similarity(set1, set2):
    """
    计算 Jaccard 相似度
    若 set2 是 set1 的子集，返回 1.0
    """
    if set2.issubset(set1):
        return 1.0
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    return intersection / union if union != 0 else 0.0

def calculate_fitness(position, target_path):
    """计算适应度（Jaccard 相似度）"""
    triggered = execute_Tr(position)
    if target_path.issubset(triggered):
        return 1.0
    intersection = len(triggered & target_path)
    union = len(triggered | target_path)
    return intersection / union if union > 0 else 0.0

# === 路径相似度矩阵 ===
def compute_path_similarity_matrix(paths):
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix

# === 路径分组 ===
def group_paths_by_similarity(paths):
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.mean(avg_sim_scores)
    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]
    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)
    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group

# === 样本生成（相似路径） ===
def generate_samples_for_similar_paths(similar_group_indices, num_total=2000, top_k=200):
    def jaccard_similarity_local(a, b):
        if not a and not b:
            return 1.0
        return len(a & b) / len(a | b) if a | b else 0.0

    def compute_robustness(state, path):
        base = execute_Tr(state)
        if not base:
            return 0.0
        rob, neighbors = 0.0, 0
        for dx in [-1, 0, 1]:
            for dy in [-1, 0, 1]:
                for dz in [-1, 0, 1]:
                    if dx == dy == dz == 0:
                        continue
                    neighbor = clip_state(state + np.array([dx, dy, dz]))
                    n_trig = execute_Tr(neighbor)
                    if not n_trig:
                        continue
                    rob += jaccard_similarity_local(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Path {path_id}\n")
            f.write("threat_count severity_level confidence_score\tScore\tSimilarity\tLengthDiff\tRobustness\n")
            for s in samples:
                count, severity, conf = s[0][0], s[0][1], s[0][2]
                f.write(f"{count} {severity:.3f} {conf:.3f}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\n")

    print("Similar path group...")
    base_dir = os.path.join(os.getcwd(), "path_samples")
    for path_idx in similar_group_indices:
        path = targetPaths[path_idx]
        samples = []
        attempts = 0
        while len(samples) < top_k and attempts < num_total * 5:
            attempts += 1
            state = generate_random_state()
            triggered = execute_Tr(state)
            if not triggered:
                continue
            sim = jaccard_similarity_local(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            score = 0.55 * sim + 0.25 * len_diff + 0.2 * rob
            samples.append((state, score, sim, len_diff, rob))
        if samples:
            samples.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples[:top_k], base_dir=base_dir)

# === 经验回放缓冲区 ===
class SharedExperienceReplay:
    def __init__(self, capacity=10000):
        self.capacity = capacity
        self.buffer = deque(maxlen=self.capacity)
        self.priorities = deque(maxlen=self.capacity)

    def append(self, experience):
        self.buffer.append(experience)
        self.priorities.append(experience[-1])

    def sample(self, batch_size, alpha=0.6):
        if len(self.buffer) == 0:
            return [], [], []
        priorities = np.array(self.priorities) ** alpha
        sum_priorities = np.sum(priorities)
        if sum_priorities == 0:
            probabilities = np.ones(len(self.buffer)) / len(self.buffer)
        else:
            probabilities = priorities / sum_priorities
        batch_indices = np.random.choice(len(self.buffer), batch_size, p=probabilities)
        batch = [self.buffer[idx] for idx in batch_indices]
        return batch, batch_indices, probabilities[batch_indices]

    def __len__(self):
        return len(self.buffer)

    def get_high_reward_samples(self, target_path, num_samples=20):
        if len(self.buffer) == 0:
            return []
        samples_with_similarity = []
        for experience in self.buffer:
            state_tensor = experience[0]
            state_tuple = tuple(state_tensor.cpu().numpy().flatten())
            triggered = execute_Tr(state_tuple)
            sim = jaccard_similarity(triggered, target_path)
            if sim >= 1.0:
                return [(state_tuple, 0, sim, triggered)]
            samples_with_similarity.append((state_tuple, 0, sim, triggered))
        samples_with_similarity.sort(key=lambda x: x[2], reverse=True)
        return samples_with_similarity[:num_samples]

def load_path_data(file_path):
    path_data = []
    if not os.path.exists(file_path):
        return path_data
    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines[2:]:
            parts = line.strip().split("\t")
            if parts:
                vals = parts[0].split()
                if len(vals) == 3:
                    state = (int(vals[0]), float(vals[1]), float(vals[2]))
                    path_data.append(state)
    return path_data

# === DQN 网络 ===
class DQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, 128)
        self.fc2 = nn.Linear(128, 64)
        self.fc3 = nn.Linear(64, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)

# === DQN Agent with PER ===
class DQNAgentWithPER:
    def __init__(self, state_dim, action_dim, replay_buffer, gamma=0.99, epsilon=1.0, epsilon_decay=0.995,
                 epsilon_min=0.1, learning_rate=0.001, alpha=0.6, beta=0.4):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer
        self.alpha = alpha
        self.beta = beta

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        """将动作索引解码为增量 (d_threat, d_severity, d_confidence)"""
        dim = action_idx // 10
        delta_idx = action_idx % 10
        if dim == 0:
            delta = DELTA_THREAT[delta_idx]
            return (delta, 0, 0)
        elif dim == 1:
            delta = DELTA_SEVERITY[delta_idx]
            return (0, delta, 0)
        elif dim == 2:
            delta = DELTA_CONFIDENCE[delta_idx]
            return (0, 0, delta)

    def act(self, state):
        if random.random() < self.epsilon:
            return random.randrange(self.action_dim)
        state = torch.tensor(state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state)
        return torch.argmax(q_values, dim=1).item()

    def store_transition(self, state, action, reward, next_state, done):
        state = torch.tensor(state, dtype=torch.float32).unsqueeze(0).to(device)
        next_state = torch.tensor(next_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state)
            next_q_values = self.target_model(next_state)
            max_next_q_values = next_q_values.max(1)[0]
            target_q_values = reward + (self.gamma * max_next_q_values * (1 - done))
            td_error = torch.abs(q_values[0][action] - target_q_values).item()
        self.replay_buffer.append((state, action, reward, next_state, done, td_error))
        return td_error

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return
        batch, batch_indices, probabilities = self.replay_buffer.sample(batch_size, alpha=self.alpha)
        states, actions, rewards, next_states, dones, _ = zip(*batch)

        states = torch.tensor(np.array([s.cpu().numpy().flatten() for s in states]), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array([ns.cpu().numpy().flatten() for ns in next_states]), dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))

        loss = nn.MSELoss()(current_q_values, target_q_values)
        self.optimizer.zero_grad()
        loss.backward()
        self.optimizer.step()

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())

# === 度量收集器 ===
class MetricsCollector:
    def __init__(self):
        self.start_time = None
        self.end_time = None
        self.total_reward = 0
        self.td_errors = []
        self.final_output_similarities = []
        self.action_improvements = []
        self.total_memory_usage = 0
        self.memory_check_count = 0
        self.step_count = 0
        self.pso_start_time = None
        self.pso_end_time = None
        self.perfect_solutions_count = 0
        self.total_paths_count = 0
        self.final_fitness_scores = []
        self.pso_convergence_iterations = []
        self.pso_perfect_solutions = []
        self.pso_reset_counts = []
        self.path_execution_times = []

    def start_training(self):
        self.start_time = time.time()
    def end_training(self):
        self.end_time = time.time()
    def start_pso_phase(self):
        self.pso_start_time = time.time()
    def end_pso_phase(self):
        self.pso_end_time = time.time()

    def record_pso_result(self, fitness, is_perfect_match, convergence_iter=None, path_id=None, method='PSO',
                          reset_count=0, execution_time=0):
        self.final_fitness_scores.append(fitness)
        self.total_paths_count += 1
        self.pso_reset_counts.append(reset_count)
        self.path_execution_times.append(execution_time)
        if is_perfect_match:
            self.perfect_solutions_count += 1
            if method == 'PSO' and convergence_iter is not None:
                self.pso_perfect_solutions.append({
                    'path_id': path_id,
                    'convergence_iteration': convergence_iter,
                    'fitness': fitness,
                    'reset_count': reset_count
                })
        if convergence_iter is not None:
            self.pso_convergence_iterations.append(convergence_iter)

    def record_step_metrics(self, reward, td_error, triggered, target_path):
        self.step_count += 1
        self.total_reward += reward
        self.td_errors.append(td_error)
        try:
            process = psutil.Process(os.getpid())
            current_memory = process.memory_info().rss / 1024 / 1024
            self.total_memory_usage += current_memory
            self.memory_check_count += 1
        except:
            pass

    def record_final_output_sample(self, triggered, target_path):
        similarity = jaccard_similarity(triggered, target_path)
        self.final_output_similarities.append(similarity)

    def record_action_improvement(self, current_reward, prev_reward):
        if prev_reward is not None:
            improvement = current_reward - prev_reward
            self.action_improvements.append(1 if improvement > 0 else 0)

# === 奖励函数 ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward

# === 训练函数（相似路径） ===
def generate_and_train_for_similar_paths(agent, similar_group, path_documents, run_metrics, episodes=500, batch_size=32,
                                         steps_per_test=5, replay_times=10, is_isolated=False):
    trained_paths = set()
    for episode in range(episodes):
        for path_idx in similar_group:
            if path_idx in trained_paths:
                continue
            file_path = os.path.join(path_documents, f"path{path_idx + 1}{'_isolated' if is_isolated else ''}.txt")
            path_data = load_path_data(file_path)
            if not path_data:
                trained_paths.add(path_idx)
                continue
            target_path = targetPaths[path_idx]
            BATCH_SIZE = 50
            N_SAMPLES = 200
            N_STEPS = 3
            N_EPOCHS = 5
            replay_count = 0

            for epoch in range(N_EPOCHS):
                print(f"  Path {path_idx + 1} - Run {epoch + 1}/{N_EPOCHS}")
                for batch_start in range(0, N_SAMPLES, BATCH_SIZE):
                    batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)
                    for test_data_idx in range(batch_start, batch_end):
                        if test_data_idx >= len(path_data):
                            break
                        state = path_data[test_data_idx]
                        prev_state = None
                        prev_triggered = None
                        prev_reward = None
                        for step in range(N_STEPS):
                            legal_actions = []
                            for a in range(agent.action_dim):
                                dx, dy, dz = agent.decode_action(a)
                                cand_next = (state[0] + dx, state[1] + dy, state[2] + dz)
                                if is_state_valid(cand_next):
                                    legal_actions.append(a)
                            if not legal_actions:
                                break
                            if random.random() < agent.epsilon:
                                action = random.choice(legal_actions)
                            else:
                                state_tensor = torch.tensor(state, dtype=torch.float32).unsqueeze(0).to(device)
                                with torch.no_grad():
                                    q_values = agent.model(state_tensor)[0]
                                action = legal_actions[torch.argmax(q_values[legal_actions]).item()]
                            dx, dy, dz = agent.decode_action(action)
                            next_state = (state[0] + dx, state[1] + dy, state[2] + dz)
                            triggered = execute_Tr(next_state)
                            reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                            done = (step == N_STEPS - 1)
                            td_error = agent.store_transition(state, action, reward, next_state, done)
                            run_metrics.record_step_metrics(reward, td_error, triggered, target_path)
                            if prev_reward is not None:
                                run_metrics.record_action_improvement(reward, prev_reward)
                            prev_state = state
                            prev_triggered = triggered
                            prev_reward = reward
                            state = next_state
                    if len(agent.replay_buffer) >= batch_size:
                        agent.train(batch_size)
                        replay_count += 1
                        if replay_count % 2 == 0:
                            agent.update_target_model()
                            print(f"     {batch_start // BATCH_SIZE + 1}/4 completed |  {replay_count}  | ")
            trained_paths.add(path_idx)
            print(f"  Path {path_idx + 1} completed\n")
        if len(trained_paths) == len(similar_group):
            break
    return agent

# === 孤立路径样本生成 ===
def generate_samples_for_isolated_paths(agent_similar, isolated_group_indices, num_total=2000, top_k=200):
    def compute_q_value(state, agent):
        state_tensor = torch.tensor(state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = agent.model(state_tensor)
        return q_values.max().item()

    def compute_robustness(state, path):
        base = execute_Tr(state)
        if not base:
            return 0.0
        rob, neighbors = 0.0, 0
        for dx in [-1, 0, 1]:
            for dy in [-1, 0, 1]:
                for dz in [-1, 0, 1]:
                    if dx == dy == dz == 0:
                        continue
                    neighbor = clip_state(state + np.array([dx, dy, dz]))
                    n_trig = execute_Tr(neighbor)
                    if not n_trig:
                        continue
                    rob += jaccard_similarity(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_isolated.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Isolated Path {path_id}\n")
            f.write("threat_count severity_level confidence_score\tScore\tSimilarity\tLengthDiff\tRobustness\tQ_normalized_complement\n")
            for s in samples:
                count, severity, conf = s[0][0], s[0][1], s[0][2]
                f.write(f"{count} {severity:.3f} {conf:.3f}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\t{s[5]:.4f}\n")

    base_dir = os.path.join(os.getcwd(), "path_samples")
    for path_idx in isolated_group_indices:
        path = targetPaths[path_idx]
        candidate_samples = []
        attempts = 0
        while len(candidate_samples) < num_total and attempts < num_total * 5:
            attempts += 1
            state = generate_random_state()
            triggered = execute_Tr(state)
            if not triggered:
                continue
            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            q_value = compute_q_value(state, agent_similar)
            candidate_samples.append((state, sim, len_diff, rob, q_value))
        if not candidate_samples:
            continue
        q_values = [sample[4] for sample in candidate_samples]
        q_min = min(q_values)
        q_max = max(q_values)
        normalized_samples = []
        for state, sim, len_diff, rob, q_value in candidate_samples:
            if q_max - q_min > 0:
                q_normalized = (q_value - q_min) / (q_max - q_min)
            else:
                q_normalized = 0.5
            q_complement = 1.0 - q_normalized
            score = 0.28 * sim + 0.1 * len_diff + 0.19 * rob + 0.43 * q_complement
            normalized_samples.append((state, score, sim, len_diff, rob, q_complement))
        normalized_samples.sort(key=lambda x: x[1], reverse=True)
        top_samples = normalized_samples[:top_k]
        save_samples(path_id=path_idx + 1, samples=top_samples, base_dir=base_dir)

# === 孤立路径训练 ===
def generate_and_train_for_isolated_paths_enhanced(agent_similar, agent_isolated, similar_group, isolated_group,
                                                   path_documents, run_metrics, episodes=500, batch_size=32,
                                                   is_isolated=True):
    trained_paths = set()
    for episode in range(episodes):
        for path_idx in isolated_group:
            if path_idx in trained_paths:
                continue
            file_path = os.path.join(path_documents, f"path{path_idx + 1}_isolated.txt")
            stage2_path_data = load_path_data(file_path)
            if not stage2_path_data:
                trained_paths.add(path_idx)
                continue
            target_path = targetPaths[path_idx]
            BATCH_SIZE = 50
            N_SAMPLES = 200
            N_STEPS = 3
            N_EPOCHS = 5
            replay_count = 0

            for epoch in range(N_EPOCHS):
                print(f"  Path {path_idx + 1} - Run {epoch + 1}/{N_EPOCHS}")
                for batch_start in range(0, N_SAMPLES, BATCH_SIZE):
                    batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)
                    for test_data_idx in range(batch_start, batch_end):
                        if test_data_idx >= len(stage2_path_data):
                            break
                        state = stage2_path_data[test_data_idx]
                        prev_state = None
                        prev_triggered = None
                        prev_reward = None
                        for step in range(N_STEPS):
                            legal_actions = []
                            for a in range(agent_isolated.action_dim):
                                dx, dy, dz = agent_isolated.decode_action(a)
                                cand_next = (state[0] + dx, state[1] + dy, state[2] + dz)
                                if is_state_valid(cand_next):
                                    legal_actions.append(a)
                            if not legal_actions:
                                break
                            if random.random() < agent_isolated.epsilon:
                                action = random.choice(legal_actions)
                            else:
                                state_tensor = torch.tensor(state, dtype=torch.float32).unsqueeze(0).to(device)
                                with torch.no_grad():
                                    q_values = agent_isolated.model(state_tensor)[0]
                                action = legal_actions[torch.argmax(q_values[legal_actions]).item()]
                            dx, dy, dz = agent_isolated.decode_action(action)
                            next_state = (state[0] + dx, state[1] + dy, state[2] + dz)
                            triggered = execute_Tr(next_state)
                            reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                            if target_path.issubset(triggered):
                                reward += 2.0
                            done = (step == N_STEPS - 1)
                            td_error = agent_isolated.store_transition(state, action, reward, next_state, done)
                            run_metrics.record_step_metrics(reward, td_error, triggered, target_path)
                            if prev_reward is not None:
                                run_metrics.record_action_improvement(reward, prev_reward)
                            prev_state = state
                            prev_triggered = triggered
                            prev_reward = reward
                            state = next_state
                    if len(agent_isolated.replay_buffer) >= batch_size:
                        agent_isolated.train(batch_size)
                        replay_count += 1
                        if replay_count % 2 == 0:
                            agent_isolated.update_target_model()
                            print(f"     {batch_start // BATCH_SIZE + 1}/4 completed |  {replay_count}  | ")
            trained_paths.add(path_idx)
            print(f"  Path {path_idx + 1} completed\n")
        if len(trained_paths) == len(isolated_group):
            break
    return agent_isolated

# === PSO 粒子 ===
class Particle:
    def __init__(self, initial_position=None):
        if initial_position is not None:
            self.position = np.array(initial_position, dtype=float)
        else:
            self.position = np.array([
                np.random.uniform(BOUNDS['threat_count'][0], BOUNDS['threat_count'][1]),
                np.random.uniform(BOUNDS['severity_level'][0], BOUNDS['severity_level'][1]),
                np.random.uniform(BOUNDS['confidence_score'][0], BOUNDS['confidence_score'][1])
            ])
        self.velocity = np.array([
            np.random.uniform(-5, 5),
            np.random.uniform(-3, 3),
            np.random.uniform(-5, 5)
        ])
        self.best_position = self.position.copy()
        self.best_fitness = 0
        self.fitness = 0

# === PSO 算法 ===
class PSO:
    def __init__(self, target_path, swarm_size=20, dqn_samples=None):
        self.target_path = target_path
        self.swarm_size = swarm_size
        self.particles = []
        self.global_best_position = None
        self.global_best_fitness = 0
        self.reset_count = 0

        if dqn_samples is not None and len(dqn_samples) > 0:
            num_direct = min(len(dqn_samples), swarm_size)
            for i in range(num_direct):
                state_tuple, reward, sim, triggered = dqn_samples[i]
                particle = Particle(initial_position=state_tuple)
                self.particles.append(particle)
            if len(self.particles) < swarm_size:
                remaining = swarm_size - len(self.particles)
                for i in range(remaining):
                    base_idx = i % len(dqn_samples)
                    state_tuple, _, _, _ = dqn_samples[base_idx]
                    perturbed = np.array(state_tuple) + np.random.randint(-10, 11, size=3)
                    perturbed = clip_state(perturbed)
                    particle = Particle(initial_position=perturbed.tolist())
                    self.particles.append(particle)
        else:
            self.particles = [Particle() for _ in range(swarm_size)]

        for particle in self.particles:
            particle.fitness = self.fitness_function(particle.position)
            if particle.fitness > particle.best_fitness:
                particle.best_fitness = particle.fitness
                particle.best_position = particle.position.copy()
            if particle.fitness > self.global_best_fitness:
                self.global_best_fitness = particle.fitness
                self.global_best_position = particle.position.copy()

    def fitness_function(self, position):
        try:
            triggered = execute_Tr(position)
            if self.target_path.issubset(triggered):
                return 1.0
            intersection = len(triggered & self.target_path)
            union = len(triggered | self.target_path)
            return intersection / union if union > 0 else 0.0
        except:
            return 0.0

    def update(self, iteration, max_iterations):
        w = 0.7
        c1 = 1.5
        c2 = 1.5
        for particle in self.particles:
            r1 = np.random.random(3)
            r2 = np.random.random(3)
            particle.velocity = (w * particle.velocity +
                                 c1 * r1 * (particle.best_position - particle.position) +
                                 c2 * r2 * (self.global_best_position - particle.position))
            max_velocity = np.array([
                (BOUNDS['threat_count'][1] - BOUNDS['threat_count'][0]) * 0.2,
                (BOUNDS['severity_level'][1] - BOUNDS['severity_level'][0]) * 0.2,
                (BOUNDS['confidence_score'][1] - BOUNDS['confidence_score'][0]) * 0.2
            ])
            particle.velocity = np.clip(particle.velocity, -max_velocity, max_velocity)
            particle.position += particle.velocity
            particle.position = clip_state(particle.position)
            particle.fitness = self.fitness_function(particle.position)
            if particle.fitness > particle.best_fitness:
                particle.best_fitness = particle.fitness
                particle.best_position = particle.position.copy()
            if particle.fitness > self.global_best_fitness:
                self.global_best_fitness = particle.fitness
                self.global_best_position = particle.position.copy()

# === Excel 导出 ===
def export_multiple_runs_to_excel(all_run_results, all_run_metrics, num_runs, filename=None):
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"DQN_PSO_{num_runs}Runs_{timestamp}.xlsx"

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    dqn_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 工作表1：运行统计
    ws1 = wb.active
    ws1.title = "运行统计"
    ws1.sheet_view.showGridLines = False
    headers = ["运行", "成功率", "成功数", "平均适应度", "平均迭代次数", "PSO时间(s)", "DQN解决"]
    col_widths = [12, 12, 12, 14, 14, 14, 12]
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, (results, run_metrics) in enumerate(zip(all_run_results, all_run_metrics), start=1):
        success_count = sum(1 for r in results if r['perfect_match'])
        success_rate = (success_count / len(targetPaths)) * 100
        avg_fitness = np.mean([r['fitness'] for r in results])
        iterations_list = []
        for r in results:
            if r.get('method') == 'DQN':
                iterations_list.append(0)
            elif r.get('convergence_iteration') is not None:
                iterations_list.append(r['convergence_iteration'])
            else:
                iterations_list.append(10000)
        avg_iterations = np.mean(iterations_list)
        dqn_solved_count = sum(1 for r in results if r.get('method') == 'DQN')
        total_time = run_metrics.pso_end_time - run_metrics.pso_start_time if run_metrics.pso_end_time else 0
        row_data = [f"Run {run_idx}", f"{success_rate:.1f}%", f"{success_count}/{len(targetPaths)}",
                    f"{avg_fitness:.4f}", f"{avg_iterations:.1f}", f"{total_time:.2f}",
                    f"{dqn_solved_count}/{len(targetPaths)}"]
        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align
            if run_idx % 2 == 0:
                cell.fill = alternate_fill
            if col == 2 and success_rate == 100.0:
                cell.fill = success_fill
            elif col == 2 and success_rate < 50.0:
                cell.fill = fail_fill
            if col == 7 and dqn_solved_count > 0:
                cell.fill = dqn_fill
    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:G{len(all_run_results)+1}"

    # 工作表2：路径统计
    ws2 = wb.create_sheet(title="路径统计")
    ws2.sheet_view.showGridLines = False
    headers2 = ["Path ID", "成功数", "成功率", "平均适应度", "平均迭代", "最小迭代", "最大迭代", "DQN解决"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14, 14]
    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    num_paths = len(targetPaths)
    for path_idx in range(num_paths):
        success_count = sum(1 for results in all_run_results if results[path_idx]['perfect_match'])
        success_rate = (success_count / num_runs) * 100
        avg_fitness = np.mean([results[path_idx]['fitness'] for results in all_run_results])
        iterations_list = []
        for results in all_run_results:
            r = results[path_idx]
            if r.get('method') == 'DQN':
                iterations_list.append(0)
            elif r.get('convergence_iteration') is not None:
                iterations_list.append(r['convergence_iteration'])
            else:
                iterations_list.append(10000)
        avg_iterations = np.mean(iterations_list) if iterations_list else 0
        min_iterations = np.min(iterations_list) if iterations_list else 0
        max_iterations = np.max(iterations_list) if iterations_list else 0
        dqn_solved_count = sum(1 for results in all_run_results
                               if results[path_idx].get('method') == 'DQN')
        row_data = [f"Path {path_idx+1}", f"{success_count}/{num_runs}", f"{success_rate:.1f}%",
                    f"{avg_fitness:.4f}", f"{avg_iterations:.1f}", f"{min_iterations}",
                    f"{max_iterations}", f"{dqn_solved_count}/{num_runs}"]
        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx+2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align
            if (path_idx+1) % 2 == 0:
                cell.fill = alternate_fill
            if col == 3 and success_rate == 100.0:
                cell.fill = success_fill
            elif col == 3 and success_rate < 50.0:
                cell.fill = fail_fill
            if col == 8 and dqn_solved_count > 0:
                cell.fill = dqn_fill
    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:H{num_paths+1}"

    # 工作表3：详细结果
    ws3 = wb.create_sheet(title="详细结果")
    ws3.sheet_view.showGridLines = False
    headers3 = ["Path", "Run", "(threat_count, severity_level, confidence_score)", "适应度", "迭代次数", "方法", "触发路径"]
    col_widths3 = [10, 10, 30, 12, 12, 12, 60]
    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(num_paths):
        for run_idx, results in enumerate(all_run_results, start=1):
            result = results[path_idx]
            best_position = result['best_position']
            fitness = result['fitness']
            triggered = result['triggered']
            method = result.get('method', 'PSO')
            if method == 'DQN':
                convergence_iter = 0
            elif result.get('convergence_iteration') is not None:
                convergence_iter = result['convergence_iteration']
            else:
                convergence_iter = 10000
            particle_str = f"({int(best_position[0])}, {best_position[1]:.3f}, {best_position[2]:.3f})"
            path_str = str(sorted(list(triggered)))
            row_data = [f"Path {path_idx+1}", f"Run {run_idx}", particle_str,
                        f"{fitness:.4f}", convergence_iter if convergence_iter < 10000 else "-",
                        method, path_str]
            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = left_align if col == 7 else center_align
                if fitness == 1.0:
                    cell.fill = dqn_fill if method == 'DQN' else success_fill
                elif fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill
            row_idx += 1
    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:G{row_idx-1}"

    # 工作表4：目标路径
    ws4 = wb.create_sheet(title="目标路径")
    ws4.sheet_view.showGridLines = False
    headers4 = ["Path ID", "目标路径", "长度"]
    col_widths4 = [12, 60, 12]
    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width
    for path_idx, target_path in enumerate(targetPaths):
        path_str = str(sorted(list(target_path)))
        row_data = [f"Path {path_idx+1}", path_str, len(target_path)]
        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx+2, column=col, value=value)
            cell.border = border
            cell.alignment = left_align if col == 2 else center_align
            if (path_idx+1) % 2 == 0:
                cell.fill = alternate_fill
    ws4.freeze_panes = 'A2'

    output_dir = os.path.join(os.getcwd(), "results")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"\n{'='*70}")
    print(f"Excel 已保存至: {filepath}")
    print(f"{'='*70}")
    print(f"包含以下工作表:")
    print(f"  1. 运行统计 - {num_runs} 次运行的汇总")
    print(f"  2. 路径统计 - 每个路径在各运行中的表现")
    print(f"  3. 详细结果 - 每次运行每个路径的具体结果")
    print(f"  4. 目标路径 - 所有目标路径集合")
    print(f"{'='*70}\n")
    return filepath

# === 单次实验 ===
def run_single_experiment(run_num, similar_group, isolated_group):
    print(f"\n{'='*100}")
    print(f"开始第 {run_num} 次运行")
    print(f"{'='*100}")

    run_metrics = MetricsCollector()
    run_metrics.start_training()

    path_documents = os.path.join(os.getcwd(), "path_samples")

    if run_num == 1:
        print("生成相似路径样本...")
        generate_samples_for_similar_paths(similar_group, num_total=2000, top_k=200)

    replay_buffer = SharedExperienceReplay(capacity=10000)
    state_dim = 3
    action_dim = 30
    agent = DQNAgentWithPER(state_dim, action_dim, replay_buffer)

    print(f"第 {run_num} 次运行 - 训练相似路径...")
    generate_and_train_for_similar_paths(agent, similar_group, path_documents, run_metrics, episodes=500, batch_size=32,
                                         is_isolated=False)

    if run_num == 1:
        print("生成孤立路径样本...")
        generate_samples_for_isolated_paths(agent, isolated_group, num_total=2000, top_k=200)

    print(f"第 {run_num} 次运行 - 训练孤立路径...")
    isolated_replay_buffer = SharedExperienceReplay(capacity=15000)
    agent_isolated = DQNAgentWithPER(state_dim, action_dim, isolated_replay_buffer)
    agent_isolated.model.load_state_dict(agent.model.state_dict())
    agent_isolated.target_model.load_state_dict(agent.model.state_dict())

    agent_isolated = generate_and_train_for_isolated_paths_enhanced(
        agent_similar=agent,
        agent_isolated=agent_isolated,
        similar_group=similar_group,
        isolated_group=isolated_group,
        path_documents=path_documents,
        run_metrics=run_metrics,
        episodes=500,
        batch_size=32,
        is_isolated=True
    )

    run_metrics.end_training()

    # 收集 DQN 高奖励样本
    dqn_best_samples = {}
    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        dqn_best_samples[path_idx] = high_reward_samples
        for state_tuple, _, sim, triggered in high_reward_samples:
            run_metrics.record_final_output_sample(triggered, target_path)
    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        dqn_best_samples[path_idx] = high_reward_samples
        for state_tuple, _, sim, triggered in high_reward_samples:
            run_metrics.record_final_output_sample(triggered, target_path)

    # PSO 阶段
    print(f"第 {run_num} 次运行 - PSO 优化")
    run_metrics.start_pso_phase()
    max_iterations = 3000
    pso_results = []

    for i, target_path in enumerate(targetPaths):
        path_start_time = time.time()
        dqn_samples_for_path = dqn_best_samples.get(i, [])

        perfect_solution_found = False
        perfect_solution_state = None
        triggered = set()

        if dqn_samples_for_path:
            for sample in dqn_samples_for_path:
                state_tuple, reward, sim, trig = sample
                if target_path.issubset(trig):
                    perfect_solution_found = True
                    perfect_solution_state = state_tuple
                    triggered = trig
                    break

        if perfect_solution_found:
            path_execution_time = time.time() - path_start_time
            pso_results.append({
                'target_path': target_path,
                'best_position': np.array(perfect_solution_state),
                'fitness': 1.0,
                'triggered': triggered,
                'perfect_match': True,
                'method': 'DQN',
                'convergence_iteration': 0,
                'early_stopped': False,
                'reset_count': 0
            })
            run_metrics.record_pso_result(1.0, True, convergence_iter=0, path_id=i+1,
                                          method='DQN', reset_count=0, execution_time=path_execution_time)
            status = "(DQN)"
        else:
            pso = PSO(target_path, swarm_size=20, dqn_samples=dqn_samples_for_path)
            converged_at_iteration = max_iterations
            early_stop = False
            for iteration in range(max_iterations):
                pso.update(iteration, max_iterations)
                if pso.global_best_fitness >= 1.0:
                    converged_at_iteration = iteration + 1
                    early_stop = True
                    break
            path_execution_time = time.time() - path_start_time
            best_position = pso.global_best_position
            triggered = execute_Tr(best_position)
            is_perfect = target_path.issubset(triggered)
            pso_results.append({
                'target_path': target_path,
                'best_position': best_position,
                'fitness': pso.global_best_fitness,
                'triggered': triggered,
                'perfect_match': is_perfect,
                'method': 'PSO',
                'convergence_iteration': converged_at_iteration,
                'early_stopped': early_stop,
                'reset_count': pso.reset_count
            })
            run_metrics.record_pso_result(
                fitness=pso.global_best_fitness,
                is_perfect_match=is_perfect,
                convergence_iter=converged_at_iteration if early_stop else None,
                path_id=i+1,
                method='PSO',
                reset_count=pso.reset_count,
                execution_time=path_execution_time
            )
            status = "(PSO)" if is_perfect else f"({pso.global_best_fitness:.3f})"
        print(f"  Path {i+1}: {status} | 耗时 {path_execution_time:.2f}s")

    run_metrics.end_pso_phase()

    success_count = sum(1 for r in pso_results if r['perfect_match'])
    success_rate = (success_count / len(targetPaths)) * 100
    pso_time = run_metrics.pso_end_time - run_metrics.pso_start_time
    print(f"\n第 {run_num} 次运行完成: 成功率 {success_rate:.1f}% ({success_count}/{len(targetPaths)}) | PSO耗时 {pso_time:.2f}s")
    return pso_results, run_metrics

# === 多次实验 ===
def run_multiple_dqn_pso_experiments(num_runs):
    print("\n" + "="*100)
    print(f"DQN-PSO (标准PSO) - {num_runs} 次运行")
    print("="*100)
    print(f"目标路径数: {len(targetPaths)} (Path 1 ~ Path {len(targetPaths)})")
    print(f"运行次数: {num_runs}")
    print(f"搜索空间: threat_count={BOUNDS['threat_count']}, severity_level={BOUNDS['severity_level']}, confidence_score={BOUNDS['confidence_score']}")
    print("="*100)

    similar_group, isolated_group = group_paths_by_similarity(targetPaths)
    print(f"相似路径组: {[idx+1 for idx in similar_group]}")
    print(f"孤立路径组: {[idx+1 for idx in isolated_group]}\n")

    all_run_results = []
    all_run_metrics = []
    total_start = time.time()

    for run_num in range(1, num_runs+1):
        pso_results, run_metrics = run_single_experiment(run_num, similar_group, isolated_group)
        all_run_results.append(pso_results)
        all_run_metrics.append(run_metrics)

    total_time = time.time() - total_start
    print(f"\n{'='*100}")
    print(f"全部 {num_runs} 次运行完成！")
    print(f"总耗时: {total_time:.2f}秒 ({total_time/60:.2f}分钟)")
    print(f"{'='*100}\n")
    return all_run_results, all_run_metrics

# === 主程序 ===
if __name__ == "__main__":
    print("="*100)
    print("DQN-PSO 威胁分析规则测试（第四个单元测试）")
    print("="*100)
    print(f"当前配置: 运行次数 = {NUM_RUNS}")
    print(f"目标路径数: {len(targetPaths)} (Path 1 ~ Path {len(targetPaths)})")
    print("="*100)

    if len(sys.argv) > 1:
        try:
            NUM_RUNS = int(sys.argv[1])
            print(f"从命令行读取运行次数: {NUM_RUNS}")
        except ValueError:
            print(f"命令行参数无效，使用默认运行次数 {NUM_RUNS}")

    print("\n开始实验...")
    all_run_results, all_run_metrics = run_multiple_dqn_pso_experiments(num_runs=NUM_RUNS)

    print("\n导出 Excel...")
    excel_filename = export_multiple_runs_to_excel(all_run_results, all_run_metrics, NUM_RUNS)

    print(f"\n程序运行完毕！")
    print(f"Excel 文件: {excel_filename}")
    print("="*100)