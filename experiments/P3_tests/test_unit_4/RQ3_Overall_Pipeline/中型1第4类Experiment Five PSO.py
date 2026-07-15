import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """安全除法"""
    if denominator == 0:
        return default
    return numerator / denominator


# === 执行威胁分析规则函数（第四个单元测试） ===
def execute_threat_analysis_rules(a):
    """
    参数 a: (threat_count, severity_level, confidence_score)
    返回: 触发的规则编号集合
    """
    threat_count, severity_level, confidence_score = int(a[0]), float(a[1]), float(a[2])
    triggered = set()
    b = {}

    # 原始规则逻辑（完整保留）
    if (threat_count >= 100) != (threat_count >= 110):
        b[0] = 1
        triggered.add(1)
    if (threat_count >= 100) != (threat_count == 100):
        b[1] = 2
        triggered.add(2)
    if (threat_count >= 100) != (threat_count >= 75):
        b[2] = 3
        triggered.add(3)

    if threat_count >= 100:
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 8.0 and confidence_score != 90):
            b[3] = 4
            triggered.add(4)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 8.0 and confidence_score >= 95):
            b[4] = 5
            triggered.add(5)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level != 8.0 and confidence_score >= 90):
            b[5] = 6
            triggered.add(6)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 9.5 and confidence_score >= 90):
            b[6] = 7
            triggered.add(7)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 9.3 and confidence_score >= 90):
            b[7] = 8
            triggered.add(8)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 8.0 and confidence_score >= 88):
            b[8] = 9
            triggered.add(9)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 8.0 and confidence_score >= 95):
            b[9] = 10
            triggered.add(10)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 8.3 and confidence_score >= 90):
            b[10] = 11
            triggered.add(11)

        if severity_level >= 8.0 and confidence_score >= 90:
            if (confidence_score >= 95) != (confidence_score >= 93):
                b[11] = 12
                triggered.add(12)
            if (confidence_score >= 95) != (confidence_score == 95):
                b[12] = 13
                triggered.add(13)
            if (confidence_score >= 95) != (confidence_score >= 97):
                b[13] = 14
                triggered.add(14)

            if (severity_level >= 9.0) != (severity_level >= 9.1):
                b[14] = 15
                triggered.add(15)
            if (severity_level >= 9.0) != (severity_level == 9.0):
                b[15] = 16
                triggered.add(16)
            if (severity_level >= 9.0) != (severity_level >= 8.0):
                b[16] = 17
                triggered.add(17)

        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level != 6.0 and confidence_score >= 80):
            b[17] = 18
            triggered.add(18)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level == 6.0 and confidence_score >= 80):
            b[18] = 19
            triggered.add(19)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score != 80):
            b[19] = 20
            triggered.add(20)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score == 80):
            b[20] = 21
            triggered.add(21)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score >= 82):
            b[21] = 22
            triggered.add(22)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score >= 88):
            b[22] = 23
            triggered.add(23)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 7.7 and confidence_score >= 82):
            b[23] = 24
            triggered.add(24)

        if (severity_level >= 4.0) != (severity_level != 4.0):
            b[24] = 25
            triggered.add(25)
        if (severity_level >= 4.0) != (severity_level >= 7.0):
            b[25] = 26
            triggered.add(26)
        if (severity_level >= 4.0) != (severity_level >= 5.0):
            b[26] = 27
            triggered.add(27)

    if (threat_count >= 50) != (threat_count >= 75):
        b[27] = 28
        triggered.add(28)
    if (threat_count >= 50) != (threat_count >= 59):
        b[28] = 29
        triggered.add(29)
    if (threat_count >= 50) != (threat_count >= 55):
        b[29] = 30
        triggered.add(30)

    elif threat_count >= 50:
        if (severity_level >= 7.0 and confidence_score >= 85) != (severity_level != 7.0 and confidence_score >= 85):
            b[30] = 31
            triggered.add(31)
        if (severity_level >= 7.0 and confidence_score >= 85) != (severity_level == 7.0 and confidence_score >= 85):
            b[31] = 32
            triggered.add(32)
        if (severity_level >= 7.0 and confidence_score >= 85) != (severity_level >= 7.0 and confidence_score != 85):
            b[32] = 33
            triggered.add(33)
        if (severity_level >= 7.0 and confidence_score >= 85) != (severity_level >= 7.0 and confidence_score == 85):
            b[33] = 34
            triggered.add(34)
        if (severity_level >= 7.0 and confidence_score >= 85) != (severity_level >= 7.0 and confidence_score >= 88):
            b[34] = 35
            triggered.add(35)

        if severity_level >= 7.0 and confidence_score >= 85:
            if (confidence_score >= 95) != (confidence_score >= 85):
                b[35] = 36
                triggered.add(36)
            if (confidence_score >= 95) != (confidence_score == 95):
                b[36] = 37
                triggered.add(37)

        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level >= 5.0 and confidence_score != 70):
            b[37] = 38
            triggered.add(38)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level >= 5.0 and confidence_score == 70):
            b[38] = 39
            triggered.add(39)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level >= 5.0 and confidence_score >= 75):
            b[39] = 40
            triggered.add(40)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level != 5.0 and confidence_score >= 70):
            b[40] = 41
            triggered.add(41)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level == 5.0 and confidence_score >= 70):
            b[41] = 42
            triggered.add(42)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level >= 5.0 and confidence_score >= 74):
            b[42] = 43
            triggered.add(43)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level >= 6.2 and confidence_score >= 70):
            b[43] = 44
            triggered.add(44)

    if (threat_count >= 20) != (threat_count != 20):
        b[44] = 45
        triggered.add(45)
    if (threat_count >= 20) != (threat_count >= 30):
        b[45] = 46
        triggered.add(46)
    if (threat_count >= 20) != (threat_count >= 25):
        b[46] = 47
        triggered.add(47)

    elif threat_count >= 20:
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score != 80):
            b[47] = 48
            triggered.add(48)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score == 80):
            b[48] = 49
            triggered.add(49)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score >= 87):
            b[49] = 50
            triggered.add(50)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level != 6.0 and confidence_score >= 80):
            b[50] = 51
            triggered.add(51)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level == 6.0 and confidence_score >= 80):
            b[51] = 52
            triggered.add(52)

        if (severity_level >= 4.0) != (severity_level != 4.0):
            b[52] = 53
            triggered.add(53)
        if (severity_level >= 4.0) != (severity_level >= 2.5):
            b[53] = 54
            triggered.add(54)

    if (threat_count >= 5) != (threat_count != 5):
        b[54] = 55
        triggered.add(55)
    if (threat_count >= 5) != (threat_count >= 7.2):
        b[55] = 56
        triggered.add(56)

    elif threat_count >= 5:
        if (severity_level >= 7.0) != (severity_level >= 8.2):
            b[56] = 57
            triggered.add(57)
        if (severity_level >= 7.0) != (severity_level == 7.0):
            b[57] = 58
            triggered.add(58)
        if (severity_level >= 7.0) != (severity_level >= 7.5):
            b[58] = 59
            triggered.add(59)

    # 严重程度维度的利益相关者通知
    if (severity_level >= 9.0) != (severity_level >= 9.4):
        b[59] = 60
        triggered.add(60)
    if (severity_level >= 9.0) != (severity_level == 9.0):
        b[60] = 61
        triggered.add(61)

    if severity_level >= 9.0:
        if (threat_count >= 50 and confidence_score >= 90) != (threat_count >= 50 and confidence_score >= 93):
            b[61] = 62
            triggered.add(62)
        if (threat_count >= 50 and confidence_score >= 90) != (threat_count >= 50 and confidence_score == 90):
            b[62] = 63
            triggered.add(63)
        if (threat_count >= 50 and confidence_score >= 90) != (threat_count >= 55 and confidence_score >= 93):
            b[63] = 64
            triggered.add(64)
        if (threat_count >= 50 and confidence_score >= 90) != (threat_count >= 66 and confidence_score >= 90):
            b[64] = 65
            triggered.add(65)
        if (threat_count >= 50 and confidence_score >= 90) != (threat_count >= 50 and confidence_score == 90):
            b[65] = 66
            triggered.add(66)

        if (threat_count >= 20 or confidence_score >= 85) != (threat_count != 20 or confidence_score >= 85):
            b[66] = 67
            triggered.add(67)
        if (threat_count >= 20 or confidence_score >= 85) != (threat_count >= 50 or confidence_score >= 85):
            b[67] = 68
            triggered.add(68)
        if (threat_count >= 20 or confidence_score >= 85) != (threat_count >= 35 or confidence_score >= 85):
            b[68] = 69
            triggered.add(69)
        if (threat_count >= 20 or confidence_score >= 85) != (threat_count >= 20 or confidence_score != 85):
            b[69] = 70
            triggered.add(70)
        if (threat_count >= 20 or confidence_score >= 85) != (threat_count >= 20 or confidence_score == 85):
            b[70] = 71
            triggered.add(71)

    if (severity_level >= 7.0) != (severity_level >= 5.5):
        b[71] = 72
        triggered.add(72)
    if (severity_level >= 7.0) != (severity_level == 7.0):
        b[72] = 73
        triggered.add(73)

    elif severity_level >= 7.0:
        if (confidence_score >= 85) != (confidence_score >= 87):
            b[73] = 74
            triggered.add(74)
        if (confidence_score >= 85) != (confidence_score == 85):
            b[74] = 75
            triggered.add(75)

    if (severity_level >= 5.0) != (severity_level != 5.0):
        b[75] = 76
        triggered.add(76)
    if (severity_level >= 5.0) != (severity_level >= 6.0):
        b[76] = 77
        triggered.add(77)

    # 置信度分数维度的行动建议
    if (confidence_score >= 95) != (confidence_score >= 97):
        b[77] = 78
        triggered.add(78)
    if (confidence_score >= 95) != (confidence_score >= 89):
        b[78] = 79
        triggered.add(79)
    if (confidence_score >= 95) != (confidence_score == 95):
        b[79] = 80
        triggered.add(80)

    if confidence_score >= 95:
        if (severity_level >= 8.0 and threat_count >= 30) != (severity_level >= 8.0 and threat_count != 30):
            b[80] = 81
            triggered.add(81)
        if (severity_level >= 8.0 and threat_count >= 30) != (severity_level >= 8.0 and threat_count == 30):
            b[81] = 82
            triggered.add(82)
        if (severity_level >= 8.0 and threat_count >= 30) != (severity_level >= 8.0 and threat_count >= 31):
            b[82] = 83
            triggered.add(83)
        if (severity_level >= 8.0 and threat_count >= 30) != (severity_level >= 8.0 and threat_count >= 45):
            b[83] = 84
            triggered.add(84)
        if (severity_level >= 8.0 and threat_count >= 30) != (severity_level == 8.0 and threat_count >= 30):
            b[84] = 85
            triggered.add(85)

        if (severity_level >= 6.0) != (severity_level >= 6.6):
            b[85] = 86
            triggered.add(86)
        if (severity_level >= 6.0) != (severity_level >= 7.3):
            b[86] = 87
            triggered.add(87)

    if (confidence_score >= 85) != (confidence_score >= 78):
        b[87] = 88
        triggered.add(88)
    if (confidence_score >= 85) != (confidence_score >= 87):
        b[88] = 89
        triggered.add(89)

    elif confidence_score >= 85:
        if (severity_level >= 7.0) != (severity_level >= 8.7):
            b[89] = 90
            triggered.add(90)
        if (severity_level >= 7.0) != (severity_level == 7.0):
            b[90] = 91
            triggered.add(91)
        if (severity_level >= 7.0) != (severity_level >= 7.2):
            b[91] = 92
            triggered.add(92)

        if (severity_level >= 5.0) != (severity_level >= 6.3):
            b[92] = 93
            triggered.add(93)
        if (severity_level >= 5.0) != (severity_level >= 3.7):
            b[93] = 94
            triggered.add(94)

    if (confidence_score >= 70) != (confidence_score >= 77):
        b[94] = 95
        triggered.add(95)
    if (confidence_score >= 70) != (confidence_score >= 86):
        b[95] = 96
        triggered.add(96)

    return triggered


# 将执行函数统一命名为 execute_Tr（供 PSO 调用）
execute_Tr = execute_threat_analysis_rules


# === 适应度函数 ===
def calculate_fitness(position: List[float], target_path: Set[int]) -> float:
    """计算 Jaccard 相似度（若完全包含目标路径则返回 1.0）"""
    generated_path = execute_Tr(position)   # position 为 [threat_count, severity_level, confidence_score]

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


# === 标准 PSO 类 ===
class BasicPSO:
    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # 修改为威胁分析的边界
        self.bounds = bounds if bounds else [(1, 120), (1, 10), (1, 100)]  # threat_count, severity_level, confidence_score
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        particles = []
        velocities = []
        for _ in range(self.n_particles):
            particle = []
            for i in range(self.dim):
                if i == 0:  # threat_count 为整数
                    particle.append(random.randint(self.bounds[i][0], self.bounds[i][1]))
                else:  # severity_level 和 confidence_score 为浮点数
                    particle.append(random.uniform(self.bounds[i][0], self.bounds[i][1]))
            particles.append(particle)
            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)
        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        new_velocity = []
        new_particle = []
        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))
            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            # threat_count 为整数，取整；其他维度保留浮点
            if i == 0:
                p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        start_time = time.time()
        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    return {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_Tr(particles[i]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()
                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

        return {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_Tr(gbest_particle),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }


# === 运行多次实验 ===
def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs} 次运行")
    print(f"{'=' * 70}")
    print(f"参数: {n_particles} 个粒子, {max_iterations} 次迭代, {len(target_paths)} 条目标路径")
    print(f"搜索空间: threat_count [1,120], severity_level [1,10], confidence_score [1,100]")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"--- 第 {run_idx}/{num_runs} 次运行 ---")
        results = {}
        pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

        for i, target_path in enumerate(target_paths):
            print(f"  Path {i+1}: ", end='')
            result = pso.optimize(target_path)
            results[i] = result
            status = "✓" if result['success'] else f"({result['best_fitness']:.3f})"
            print(f"{status} | 耗时 {result['time']:.2f}s | 迭代 {result['iterations']}")

        all_results.append(results)
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f"  本轮成功: {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start
    print(f"{'=' * 70}")
    print(f"全部 {num_runs} 次运行完成 | 总耗时 {total_time:.2f}s")
    print(f"{'=' * 70}\n")
    return all_results


# === 导出 Excel ===
def export_to_excel(all_results, target_paths, filename=None):
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Threat_Results_{timestamp}.xlsx"

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 工作表1：运行统计
    ws1 = wb.active
    ws1.title = "运行统计"
    ws1.sheet_view.showGridLines = False
    headers = ["运行", "成功率", "成功数", "平均适应度", "平均迭代次数", "PSO时间(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = sum(results[i]['time'] for i in range(len(target_paths)))

        row_data = [
            f"Run {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx+1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align
            if run_idx % 2 == 0:
                cell.fill = alternate_fill
            if col == 2 and success_rate == 100.0:
                cell.fill = success_fill
            elif col == 2 and success_rate < 50.0:
                cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results)+1}"

    # 工作表2：路径统计
    ws2 = wb.create_sheet(title="路径统计")
    ws2.sheet_view.showGridLines = False
    headers2 = ["Path ID", "成功数", "成功率", "平均适应度", "平均迭代", "最小迭代", "最大迭代"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]
    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])
        iters = [r[path_idx]['iterations'] for r in all_results]
        row_data = [
            f"Path {path_idx+1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{np.mean(iters):.1f}",
            f"{np.min(iters)}",
            f"{np.max(iters)}"
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx+2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align
            if (path_idx+1) % 2 == 0:
                cell.fill = alternate_fill
            if col == 3 and success_rate == 100.0:
                cell.fill = success_fill
            elif col == 3 and success_rate < 50.0:
                cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths)+1}"

    # 工作表3：详细结果
    ws3 = wb.create_sheet(title="详细结果")
    ws3.sheet_view.showGridLines = False
    headers3 = ["Path", "Run", "(threat_count, severity_level, confidence_score)", "适应度", "迭代次数", "触发路径"]
    col_widths3 = [10, 10, 30, 12, 12, 55]
    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            result = results[path_idx]
            pos = result['best_particle']
            fitness = result['best_fitness']
            path_set = result['best_path']
            iters = result['iterations']
            row_data = [
                f"Path {path_idx+1}",
                f"Run {run_idx}",
                f"({int(pos[0])}, {pos[1]:.3f}, {pos[2]:.3f})",
                f"{fitness:.4f}",
                iters,
                str(sorted(list(path_set)))
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = left_align if col == 6 else center_align
                if fitness == 1.0:
                    cell.fill = success_fill
                elif fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill
            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx-1}"

    # 工作表4：目标路径
    ws4 = wb.create_sheet(title="目标路径")
    ws4.sheet_view.showGridLines = False
    headers4 = ["Path ID", "目标路径", "长度"]
    col_widths4 = [12, 60, 12]
    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        row_data = [
            f"Path {path_idx+1}",
            str(sorted(list(target_path))),
            len(target_path)
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx+2, column=col, value=value)
            cell.border = border
            cell.alignment = left_align if col == 2 else center_align
            if (path_idx+1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)
    print(f"\n{'=' * 70}")
    print(f"Excel 已保存: {filename}")
    print(f"{'=' * 70}")
    return filename


# === 主函数 ===
def main():
    # === 目标路径组（威胁分析，26条） ===
    targetPaths = [
        # A1
        {6, 7, 8, 11, 13, 17, 19, 21, 28, 29, 30, 32, 34, 37, 39, 42, 49, 52, 57, 58, 63, 64, 65, 66, 73, 75, 80, 82,
         90, 91},
        # A2
        {5, 7, 8, 10, 15, 19, 21, 28, 29, 32, 34, 36, 39, 42, 49, 52, 58, 60, 62, 63, 64, 65, 66, 73, 75, 79, 82, 85,
         91},
        # A3
        {6, 7, 8, 11, 14, 17, 19, 21, 28, 29, 30, 32, 34, 39, 42, 49, 52, 57, 58, 63, 64, 65, 66, 73, 75, 78, 82, 90,
         91},
        # A4
        {4, 5, 7, 8, 10, 15, 19, 21, 28, 29, 30, 32, 34, 36, 39, 42, 49, 52, 58, 60, 62, 64, 65, 73, 75, 79, 82, 85,
         91},
        # A5
        {5, 6, 7, 8, 10, 11, 12, 17, 19, 21, 28, 32, 34, 36, 39, 42, 49, 52, 57, 58, 63, 65, 66, 73, 75, 79, 82, 90,
         91},
        # A6
        {5, 10, 16, 19, 21, 28, 29, 30, 32, 34, 36, 39, 42, 49, 52, 58, 61, 62, 63, 64, 65, 66, 73, 75, 79, 82, 85, 91},
        # A7
        {1, 2, 6, 7, 8, 11, 13, 14, 17, 19, 21, 32, 34, 37, 39, 42, 49, 52, 57, 58, 63, 66, 73, 75, 78, 80, 82, 90, 91},
        # A8
        {5, 6, 7, 8, 10, 11, 12, 17, 19, 21, 32, 34, 36, 39, 42, 45, 49, 52, 55, 57, 58, 71, 73, 75, 79, 81, 90, 91},
        # A9
        {3, 6, 7, 8, 11, 13, 14, 17, 19, 21, 32, 34, 37, 39, 42, 49, 52, 57, 58, 63, 66, 73, 75, 78, 80, 82, 90, 91},
        # A10
        {4, 5, 6, 7, 8, 10, 11, 17, 19, 21, 32, 34, 36, 39, 42, 45, 46, 47, 49, 52, 57, 58, 73, 75, 79, 81, 90, 91},
        # A11
        {6, 7, 8, 11, 13, 17, 19, 21, 32, 34, 37, 39, 42, 45, 49, 52, 55, 56, 57, 58, 71, 73, 75, 80, 81, 90, 91},
        # A12
        {4, 15, 19, 21, 23, 28, 29, 30, 32, 34, 35, 36, 39, 42, 49, 50, 52, 58, 60, 73, 74, 75, 82, 85, 89, 91},
        # A13
        {4, 17, 19, 21, 23, 32, 33, 35, 36, 39, 42, 45, 46, 47, 49, 50, 52, 57, 58, 73, 74, 81, 89, 90, 91, 96},
        # A14
        {4, 17, 19, 21, 23, 32, 33, 35, 36, 39, 42, 45, 49, 50, 52, 56, 57, 58, 70, 73, 74, 81, 89, 90, 91, 96},
        # A15
        {4, 17, 19, 21, 23, 33, 39, 42, 45, 46, 47, 49, 50, 52, 57, 58, 67, 68, 69, 73, 81, 88, 90, 91, 96},
        # A16
        {4, 17, 19, 21, 22, 23, 24, 33, 39, 42, 45, 49, 50, 52, 55, 57, 58, 67, 70, 73, 81, 88, 90, 91, 96},
        # A17
        {6, 13, 14, 18, 21, 24, 26, 28, 31, 37, 39, 42, 44, 49, 51, 63, 65, 66, 72, 75, 78, 80, 86, 87, 93},
        # A18
        {6, 13, 14, 19, 21, 24, 28, 31, 34, 37, 39, 42, 49, 52, 57, 59, 63, 65, 66, 75, 78, 80, 87, 90, 92},
        # A19
        {5, 6, 7, 8, 10, 11, 17, 19, 21, 32, 34, 36, 39, 42, 49, 52, 57, 58, 73, 75, 79, 82, 84, 90, 91},
        # A20
        {4, 16, 19, 20, 22, 23, 24, 28, 29, 30, 33, 39, 42, 48, 50, 52, 58, 61, 73, 82, 85, 88, 91, 96},
        # A21
        {6, 13, 14, 18, 26, 28, 29, 30, 31, 37, 39, 41, 44, 51, 63, 64, 65, 66, 75, 76, 77, 78, 80, 93},
        # A22
        {6, 7, 8, 11, 14, 17, 19, 21, 32, 34, 39, 42, 49, 52, 57, 58, 73, 75, 78, 81, 83, 84, 90, 91},
        # A23
        {4, 9, 17, 19, 21, 32, 34, 36, 39, 42, 45, 46, 47, 49, 52, 57, 58, 73, 75, 79, 81, 90, 91},
        # A24
        {4, 17, 20, 33, 38, 40, 42, 43, 45, 48, 55, 56, 57, 58, 67, 70, 73, 81, 90, 91, 95, 96},
        # A25
        {6, 18, 25, 26, 27, 28, 29, 30, 31, 36, 41, 51, 53, 62, 63, 64, 65, 66, 75, 76, 79, 94},
        # A26
        {6, 12, 18, 25, 28, 29, 30, 31, 36, 41, 51, 53, 54, 63, 64, 65, 66, 75, 76, 79}
    ]

    print("=" * 70)
    print("baseline PSO - 威胁分析规则单元测试（第四个单元测试）")
    print("=" * 70)

    all_results = run_multiple_experiments(
        targetPaths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, targetPaths)

    print("Program completed")


if __name__ == "__main__":
    main()