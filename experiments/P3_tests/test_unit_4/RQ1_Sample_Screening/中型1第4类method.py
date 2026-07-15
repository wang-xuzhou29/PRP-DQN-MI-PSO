import os
import random
from collections import deque
import numpy as np
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# ==================== 新的全局范围配置 ====================
THREAT_COUNT_MIN, THREAT_COUNT_MAX = 1, 120
SEVERITY_LEVEL_MIN, SEVERITY_LEVEL_MAX = 1.0, 10.0
CONFIDENCE_SCORE_MIN, CONFIDENCE_SCORE_MAX = 1.0, 100.0

# 邻域扰动步长
STEP_THREAT = 1
STEP_SEVERITY = 0.1
STEP_CONFIDENCE = 1.0  # 可根据需要调整

def generate_input():
    """生成随机输入（兼容旧接口）"""
    return [
        random.randint(THREAT_COUNT_MIN, THREAT_COUNT_MAX),
        random.uniform(SEVERITY_LEVEL_MIN, SEVERITY_LEVEL_MAX),
        random.uniform(CONFIDENCE_SCORE_MIN, CONFIDENCE_SCORE_MAX)
    ]

# === 执行威胁分析规则函数 ===
def execute_threat_analysis_rules(a):
    """
    威胁分析规则函数
    参数a: 包含3个元素的元组或数组，分别对应 threat_count, severity_level, confidence_score
    返回: 触发的规则编号集合
    """
    threat_count, severity_level, confidence_score = int(a[0]), float(a[1]), float(a[2])
    triggered = set()
    b = {}

    # 以下为原始规则逻辑，未作任何修改
    if (threat_count >= 100) != (threat_count >= 110):
        b[0] = 1
        triggered.add(1)
    if (threat_count >= 100) != (threat_count == 100):
        b[1] = 2
        triggered.add(2)
    if (threat_count >= 100) != (threat_count >= 75):
        b[2] = 3
        triggered.add(3)

    if threat_count >= 100:
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 8.0 and confidence_score != 90):
            b[3] = 4
            triggered.add(4)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 8.0 and confidence_score >= 95):
            b[4] = 5
            triggered.add(5)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level != 8.0 and confidence_score >= 90):
            b[5] = 6
            triggered.add(6)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 9.5 and confidence_score >= 90):
            b[6] = 7
            triggered.add(7)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 9.3 and confidence_score >= 90):
            b[7] = 8
            triggered.add(8)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 8.0 and confidence_score >= 88):
            b[8] = 9
            triggered.add(9)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 8.0 and confidence_score >= 95):
            b[9] = 10
            triggered.add(10)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 8.3 and confidence_score >= 90):
            b[10] = 11
            triggered.add(11)

        if severity_level >= 8.0 and confidence_score >= 90:
            if (confidence_score >= 95) != (confidence_score >= 93):
                b[11] = 12
                triggered.add(12)
            if (confidence_score >= 95) != (confidence_score == 95):
                b[12] = 13
                triggered.add(13)
            if (confidence_score >= 95) != (confidence_score >= 97):
                b[13] = 14
                triggered.add(14)

            if (severity_level >= 9.0) != (severity_level >= 9.1):
                b[14] = 15
                triggered.add(15)
            if (severity_level >= 9.0) != (severity_level == 9.0):
                b[15] = 16
                triggered.add(16)
            if (severity_level >= 9.0) != (severity_level >= 8.0):
                b[16] = 17
                triggered.add(17)

        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level != 6.0 and confidence_score >= 80):
            b[17] = 18
            triggered.add(18)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level == 6.0 and confidence_score >= 80):
            b[18] = 19
            triggered.add(19)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score != 80):
            b[19] = 20
            triggered.add(20)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score == 80):
            b[20] = 21
            triggered.add(21)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score >= 82):
            b[21] = 22
            triggered.add(22)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score >= 88):
            b[22] = 23
            triggered.add(23)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 7.7 and confidence_score >= 82):
            b[23] = 24
            triggered.add(24)

        if (severity_level >= 4.0) != (severity_level != 4.0):
            b[24] = 25
            triggered.add(25)
        if (severity_level >= 4.0) != (severity_level >= 7.0):
            b[25] = 26
            triggered.add(26)
        if (severity_level >= 4.0) != (severity_level >= 5.0):
            b[26] = 27
            triggered.add(27)

    if (threat_count >= 50) != (threat_count >= 75):
        b[27] = 28
        triggered.add(28)
    if (threat_count >= 50) != (threat_count >= 59):
        b[28] = 29
        triggered.add(29)
    if (threat_count >= 50) != (threat_count >= 55):
        b[29] = 30
        triggered.add(30)

    elif threat_count >= 50:
        if (severity_level >= 7.0 and confidence_score >= 85) != (severity_level != 7.0 and confidence_score >= 85):
            b[30] = 31
            triggered.add(31)
        if (severity_level >= 7.0 and confidence_score >= 85) != (severity_level == 7.0 and confidence_score >= 85):
            b[31] = 32
            triggered.add(32)
        if (severity_level >= 7.0 and confidence_score >= 85) != (severity_level >= 7.0 and confidence_score != 85):
            b[32] = 33
            triggered.add(33)
        if (severity_level >= 7.0 and confidence_score >= 85) != (severity_level >= 7.0 and confidence_score == 85):
            b[33] = 34
            triggered.add(34)
        if (severity_level >= 7.0 and confidence_score >= 85) != (severity_level >= 7.0 and confidence_score >= 88):
            b[34] = 35
            triggered.add(35)

        if severity_level >= 7.0 and confidence_score >= 85:
            if (confidence_score >= 95) != (confidence_score >= 85):
                b[35] = 36
                triggered.add(36)
            if (confidence_score >= 95) != (confidence_score == 95):
                b[36] = 37
                triggered.add(37)

        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level >= 5.0 and confidence_score != 70):
            b[37] = 38
            triggered.add(38)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level >= 5.0 and confidence_score == 70):
            b[38] = 39
            triggered.add(39)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level >= 5.0 and confidence_score >= 75):
            b[39] = 40
            triggered.add(40)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level != 5.0 and confidence_score >= 70):
            b[40] = 41
            triggered.add(41)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level == 5.0 and confidence_score >= 70):
            b[41] = 42
            triggered.add(42)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level >= 5.0 and confidence_score >= 74):
            b[42] = 43
            triggered.add(43)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level >= 6.2 and confidence_score >= 70):
            b[43] = 44
            triggered.add(44)

    if (threat_count >= 20) != (threat_count != 20):
        b[44] = 45
        triggered.add(45)
    if (threat_count >= 20) != (threat_count >= 30):
        b[45] = 46
        triggered.add(46)
    if (threat_count >= 20) != (threat_count >= 25):
        b[46] = 47
        triggered.add(47)

    elif threat_count >= 20:
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score != 80):
            b[47] = 48
            triggered.add(48)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score == 80):
            b[48] = 49
            triggered.add(49)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score >= 87):
            b[49] = 50
            triggered.add(50)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level != 6.0 and confidence_score >= 80):
            b[50] = 51
            triggered.add(51)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level == 6.0 and confidence_score >= 80):
            b[51] = 52
            triggered.add(52)

        if (severity_level >= 4.0) != (severity_level != 4.0):
            b[52] = 53
            triggered.add(53)
        if (severity_level >= 4.0) != (severity_level >= 2.5):
            b[53] = 54
            triggered.add(54)

    if (threat_count >= 5) != (threat_count != 5):
        b[54] = 55
        triggered.add(55)
    if (threat_count >= 5) != (threat_count >= 7.2):
        b[55] = 56
        triggered.add(56)

    elif threat_count >= 5:
        if (severity_level >= 7.0) != (severity_level >= 8.2):
            b[56] = 57
            triggered.add(57)
        if (severity_level >= 7.0) != (severity_level == 7.0):
            b[57] = 58
            triggered.add(58)
        if (severity_level >= 7.0) != (severity_level >= 7.5):
            b[58] = 59
            triggered.add(59)

    # 严重程度维度的利益相关者通知
    if (severity_level >= 9.0) != (severity_level >= 9.4):
        b[59] = 60
        triggered.add(60)
    if (severity_level >= 9.0) != (severity_level == 9.0):
        b[60] = 61
        triggered.add(61)

    if severity_level >= 9.0:
        if (threat_count >= 50 and confidence_score >= 90) != (threat_count >= 50 and confidence_score >= 93):
            b[61] = 62
            triggered.add(62)
        if (threat_count >= 50 and confidence_score >= 90) != (threat_count >= 50 and confidence_score == 90):
            b[62] = 63
            triggered.add(63)
        if (threat_count >= 50 and confidence_score >= 90) != (threat_count >= 55 and confidence_score >= 93):
            b[63] = 64
            triggered.add(64)
        if (threat_count >= 50 and confidence_score >= 90) != (threat_count >= 66 and confidence_score >= 90):
            b[64] = 65
            triggered.add(65)
        if (threat_count >= 50 and confidence_score >= 90) != (threat_count >= 50 and confidence_score == 90):
            b[65] = 66
            triggered.add(66)

        if (threat_count >= 20 or confidence_score >= 85) != (threat_count != 20 or confidence_score >= 85):
            b[66] = 67
            triggered.add(67)
        if (threat_count >= 20 or confidence_score >= 85) != (threat_count >= 50 or confidence_score >= 85):
            b[67] = 68
            triggered.add(68)
        if (threat_count >= 20 or confidence_score >= 85) != (threat_count >= 35 or confidence_score >= 85):
            b[68] = 69
            triggered.add(69)
        if (threat_count >= 20 or confidence_score >= 85) != (threat_count >= 20 or confidence_score != 85):
            b[69] = 70
            triggered.add(70)
        if (threat_count >= 20 or confidence_score >= 85) != (threat_count >= 20 or confidence_score == 85):
            b[70] = 71
            triggered.add(71)

    if (severity_level >= 7.0) != (severity_level >= 5.5):
        b[71] = 72
        triggered.add(72)
    if (severity_level >= 7.0) != (severity_level == 7.0):
        b[72] = 73
        triggered.add(73)

    elif severity_level >= 7.0:
        if (confidence_score >= 85) != (confidence_score >= 87):
            b[73] = 74
            triggered.add(74)
        if (confidence_score >= 85) != (confidence_score == 85):
            b[74] = 75
            triggered.add(75)

    if (severity_level >= 5.0) != (severity_level != 5.0):
        b[75] = 76
        triggered.add(76)
    if (severity_level >= 5.0) != (severity_level >= 6.0):
        b[76] = 77
        triggered.add(77)

    # 置信度分数维度的行动建议
    if (confidence_score >= 95) != (confidence_score >= 97):
        b[77] = 78
        triggered.add(78)
    if (confidence_score >= 95) != (confidence_score >= 89):
        b[78] = 79
        triggered.add(79)
    if (confidence_score >= 95) != (confidence_score == 95):
        b[79] = 80
        triggered.add(80)

    if confidence_score >= 95:
        if (severity_level >= 8.0 and threat_count >= 30) != (severity_level >= 8.0 and threat_count != 30):
            b[80] = 81
            triggered.add(81)
        if (severity_level >= 8.0 and threat_count >= 30) != (severity_level >= 8.0 and threat_count == 30):
            b[81] = 82
            triggered.add(82)
        if (severity_level >= 8.0 and threat_count >= 30) != (severity_level >= 8.0 and threat_count >= 31):
            b[82] = 83
            triggered.add(83)
        if (severity_level >= 8.0 and threat_count >= 30) != (severity_level >= 8.0 and threat_count >= 45):
            b[83] = 84
            triggered.add(84)
        if (severity_level >= 8.0 and threat_count >= 30) != (severity_level == 8.0 and threat_count >= 30):
            b[84] = 85
            triggered.add(85)

        if (severity_level >= 6.0) != (severity_level >= 6.6):
            b[85] = 86
            triggered.add(86)
        if (severity_level >= 6.0) != (severity_level >= 7.3):
            b[86] = 87
            triggered.add(87)

    if (confidence_score >= 85) != (confidence_score >= 78):
        b[87] = 88
        triggered.add(88)
    if (confidence_score >= 85) != (confidence_score >= 87):
        b[88] = 89
        triggered.add(89)

    elif confidence_score >= 85:
        if (severity_level >= 7.0) != (severity_level >= 8.7):
            b[89] = 90
            triggered.add(90)
        if (severity_level >= 7.0) != (severity_level == 7.0):
            b[90] = 91
            triggered.add(91)
        if (severity_level >= 7.0) != (severity_level >= 7.2):
            b[91] = 92
            triggered.add(92)

        if (severity_level >= 5.0) != (severity_level >= 6.3):
            b[92] = 93
            triggered.add(93)
        if (severity_level >= 5.0) != (severity_level >= 3.7):
            b[93] = 94
            triggered.add(94)

    if (confidence_score >= 70) != (confidence_score >= 77):
        b[94] = 95
        triggered.add(95)
    if (confidence_score >= 70) != (confidence_score >= 86):
        b[95] = 96
        triggered.add(96)

    return triggered

# 为兼容原有代码，设置别名
execute_Tr = execute_threat_analysis_rules

def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0

# === 目标路径组 ===
targetPaths = [
    # A1
    {6, 7, 8, 11, 13, 17, 19, 21, 28, 29, 30, 32, 34, 37, 39, 42, 49, 52, 57, 58, 63, 64, 65, 66, 73, 75, 80, 82, 90, 91},
    # A2
    {5, 7, 8, 10, 15, 19, 21, 28, 29, 32, 34, 36, 39, 42, 49, 52, 58, 60, 62, 63, 64, 65, 66, 73, 75, 79, 82, 85, 91},
    # A3
    {6, 7, 8, 11, 14, 17, 19, 21, 28, 29, 30, 32, 34, 39, 42, 49, 52, 57, 58, 63, 64, 65, 66, 73, 75, 78, 82, 90, 91},
    # A4
    {4, 5, 7, 8, 10, 15, 19, 21, 28, 29, 30, 32, 34, 36, 39, 42, 49, 52, 58, 60, 62, 64, 65, 73, 75, 79, 82, 85, 91},
    # A5
    {5, 6, 7, 8, 10, 11, 12, 17, 19, 21, 28, 32, 34, 36, 39, 42, 49, 52, 57, 58, 63, 65, 66, 73, 75, 79, 82, 90, 91},
    # A6
    {5, 10, 16, 19, 21, 28, 29, 30, 32, 34, 36, 39, 42, 49, 52, 58, 61, 62, 63, 64, 65, 66, 73, 75, 79, 82, 85, 91},
    # A7
    {1, 2, 6, 7, 8, 11, 13, 14, 17, 19, 21, 32, 34, 37, 39, 42, 49, 52, 57, 58, 63, 66, 73, 75, 78, 80, 82, 90, 91},
    # A8
    {5, 6, 7, 8, 10, 11, 12, 17, 19, 21, 32, 34, 36, 39, 42, 45, 49, 52, 55, 57, 58, 71, 73, 75, 79, 81, 90, 91},
    # A9
    {3, 6, 7, 8, 11, 13, 14, 17, 19, 21, 32, 34, 37, 39, 42, 49, 52, 57, 58, 63, 66, 73, 75, 78, 80, 82, 90, 91},
    # A10
    {4, 5, 6, 7, 8, 10, 11, 17, 19, 21, 32, 34, 36, 39, 42, 45, 46, 47, 49, 52, 57, 58, 73, 75, 79, 81, 90, 91},
    # A11
    {6, 7, 8, 11, 13, 17, 19, 21, 32, 34, 37, 39, 42, 45, 49, 52, 55, 56, 57, 58, 71, 73, 75, 80, 81, 90, 91},
    # A12
    {4, 15, 19, 21, 23, 28, 29, 30, 32, 34, 35, 36, 39, 42, 49, 50, 52, 58, 60, 73, 74, 75, 82, 85, 89, 91},
    # A13
    {4, 17, 19, 21, 23, 32, 33, 35, 36, 39, 42, 45, 46, 47, 49, 50, 52, 57, 58, 73, 74, 81, 89, 90, 91, 96},
    # A14
    {4, 17, 19, 21, 23, 32, 33, 35, 36, 39, 42, 45, 49, 50, 52, 56, 57, 58, 70, 73, 74, 81, 89, 90, 91, 96},
    # A15
    {4, 17, 19, 21, 23, 33, 39, 42, 45, 46, 47, 49, 50, 52, 57, 58, 67, 68, 69, 73, 81, 88, 90, 91, 96},
    # A16
    {4, 17, 19, 21, 22, 23, 24, 33, 39, 42, 45, 49, 50, 52, 55, 57, 58, 67, 70, 73, 81, 88, 90, 91, 96},
    # A17
    {6, 13, 14, 18, 21, 24, 26, 28, 31, 37, 39, 42, 44, 49, 51, 63, 65, 66, 72, 75, 78, 80, 86, 87, 93},
    # A18
    {6, 13, 14, 19, 21, 24, 28, 31, 34, 37, 39, 42, 49, 52, 57, 59, 63, 65, 66, 75, 78, 80, 87, 90, 92},
    # A19
    {5, 6, 7, 8, 10, 11, 17, 19, 21, 32, 34, 36, 39, 42, 49, 52, 57, 58, 73, 75, 79, 82, 84, 90, 91},
    # A20
    {4, 16, 19, 20, 22, 23, 24, 28, 29, 30, 33, 39, 42, 48, 50, 52, 58, 61, 73, 82, 85, 88, 91, 96},
    # A21
    {6, 13, 14, 18, 26, 28, 29, 30, 31, 37, 39, 41, 44, 51, 63, 64, 65, 66, 75, 76, 77, 78, 80, 93},
    # A22
    {6, 7, 8, 11, 14, 17, 19, 21, 32, 34, 39, 42, 49, 52, 57, 58, 73, 75, 78, 81, 83, 84, 90, 91},
    # A23
    {4, 9, 17, 19, 21, 32, 34, 36, 39, 42, 45, 46, 47, 49, 52, 57, 58, 73, 75, 79, 81, 90, 91},
    # A24
    {4, 17, 20, 33, 38, 40, 42, 43, 45, 48, 55, 56, 57, 58, 67, 70, 73, 81, 90, 91, 95, 96},
    # A25
    {6, 18, 25, 26, 27, 28, 29, 30, 31, 36, 41, 51, 53, 62, 63, 64, 65, 66, 75, 76, 79, 94},
    # A26
    {6, 12, 18, 25, 28, 29, 30, 31, 36, 41, 51, 53, 54, 63, 64, 65, 66, 75, 76, 79}
]

class ExperimentConfig:
    def __init__(self):
        self.num_total_samples = 2000
        self.top_k_samples = 200
        self.num_runs = 3
        self.test_paths = list(range(len(targetPaths)))

    STRATEGIES = {
        'random': None,
        'equal_weight': [0.33, 0.33, 0.33],
        'weighted': [0.3, 0.3, 0.40]
    }

def compute_robustness(state, path):
    base = execute_Tr(*state)
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = np.array([
                    np.clip(state[0] + dx * STEP_THREAT, THREAT_COUNT_MIN, THREAT_COUNT_MAX),
                    np.clip(state[1] + dy * STEP_SEVERITY, SEVERITY_LEVEL_MIN, SEVERITY_LEVEL_MAX),
                    np.clip(state[2] + dz * STEP_CONFIDENCE, CONFIDENCE_SCORE_MIN, CONFIDENCE_SCORE_MAX)
                ])
                n_trig = execute_Tr(*neighbor)
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0

def generate_candidate_samples(target_path_idx, sample_count=1000):
    target_path = targetPaths[target_path_idx]
    samples = []
    attempts = 0

    while len(samples) < sample_count and attempts < sample_count * 10:
        attempts += 1
        state = np.array([
            random.randint(THREAT_COUNT_MIN, THREAT_COUNT_MAX),
            random.uniform(SEVERITY_LEVEL_MIN, SEVERITY_LEVEL_MAX),
            random.uniform(CONFIDENCE_SCORE_MIN, CONFIDENCE_SCORE_MAX)
        ])
        triggered = execute_Tr(*state)

        if not triggered:
            continue

        sim = jaccard_similarity(triggered, target_path)
        len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
        rob = compute_robustness(state, target_path)

        sample_data = {
            'state': tuple(state),
            'similarity': sim,
            'length_diff': len_diff,
            'robustness': rob,
            'triggered': triggered
        }
        samples.append(sample_data)

    return samples

def apply_strategy_screening(candidate_samples, strategy_name, weights, config):
    if strategy_name == 'random':
        selected_samples = random.sample(candidate_samples, min(config.top_k_samples, len(candidate_samples)))
    else:
        samples_with_scores = []
        for sample in candidate_samples:
            score = weights[0] * sample['similarity'] + weights[1] * sample['length_diff'] + weights[2] * sample['robustness']
            sample_copy = sample.copy()
            sample_copy['score'] = score
            samples_with_scores.append(sample_copy)

        samples_with_scores.sort(key=lambda x: x['score'], reverse=True)
        selected_samples = samples_with_scores[:config.top_k_samples]

    return selected_samples

def generate_samples_with_strategy(target_path_idx, strategy_name, weights, config, shared_candidates=None):
    if strategy_name == 'random':
        target_path = targetPaths[target_path_idx]
        samples = []
        attempts = 0

        while len(samples) < config.top_k_samples and attempts < config.top_k_samples * 10:
            attempts += 1
            state = np.array([
                random.randint(THREAT_COUNT_MIN, THREAT_COUNT_MAX),
                random.uniform(SEVERITY_LEVEL_MIN, SEVERITY_LEVEL_MAX),
                random.uniform(CONFIDENCE_SCORE_MIN, CONFIDENCE_SCORE_MAX)
            ])
            triggered = execute_Tr(*state)

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, target_path)
            len_diff = 1 - abs(len(triggered) - len(target_path)) / max(len(triggered), len(target_path))
            rob = compute_robustness(state, target_path)

            sample_data = {
                'state': tuple(state),
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            }
            samples.append(sample_data)

        random.shuffle(samples)
        return samples[:config.top_k_samples]

    else:
        if shared_candidates is None:
            raise ValueError("Equal weight and weighted strategies require shared candidate samples")
        return apply_strategy_screening(shared_candidates, strategy_name, weights, config)

def run_single_experiment(config):
    results = {strategy: {} for strategy in config.STRATEGIES.keys()}
    shared_candidates = {}

    for path_idx in config.test_paths:
        candidates = generate_candidate_samples(path_idx, 1000)
        shared_candidates[path_idx] = candidates

    for strategy_name, weights in config.STRATEGIES.items():
        strategy_results = {}

        for path_idx in config.test_paths:
            if strategy_name == 'random':
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config)
            else:
                samples = generate_samples_with_strategy(path_idx, strategy_name, weights, config,
                                                         shared_candidates[path_idx])

            strategy_results[path_idx] = samples

        results[strategy_name] = strategy_results

    return results

def analyze_fitness_values(results, config):
    analysis_results = {}

    for strategy_name in config.STRATEGIES.keys():
        all_similarities = []
        all_length_diffs = []
        all_robustness = []

        for path_idx in config.test_paths:
            path_samples = results[strategy_name][path_idx]
            all_similarities.extend([s['similarity'] for s in path_samples])
            all_length_diffs.extend([s['length_diff'] for s in path_samples])
            all_robustness.extend([s['robustness'] for s in path_samples])

        analysis = {
            'mean_similarity': np.mean(all_similarities),
            'mean_length_diff': np.mean(all_length_diffs),
            'mean_robustness': np.mean(all_robustness),
            'total_samples': len(all_similarities)
        }

        all_scores = []

        if strategy_name == 'random':
            equal_weights = config.STRATEGIES['equal_weight']
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                for sample in path_samples:
                    score = (equal_weights[0] * sample['similarity'] +
                             equal_weights[1] * sample['length_diff'] +
                             equal_weights[2] * sample['robustness'])
                    all_scores.append(score)
        else:
            for path_idx in config.test_paths:
                path_samples = results[strategy_name][path_idx]
                all_scores.extend([s['score'] for s in path_samples])

        scores_array = np.array(all_scores)

        high_score_ratio = np.mean(scores_array > 0.8)
        medium_score_ratio = np.mean((scores_array >= 0.5) & (scores_array <= 0.8))
        low_score_ratio = np.mean(scores_array < 0.5)

        analysis.update({
            'mean_score': np.mean(scores_array),
            'max_score': np.max(scores_array),
            'std_score': np.std(scores_array),
            'high_score_ratio': high_score_ratio,
            'medium_score_ratio': medium_score_ratio,
            'low_score_ratio': low_score_ratio
        })

        analysis_results[strategy_name] = analysis

    return analysis_results

def compare_strategies(analysis_results):
    strategies = list(analysis_results.keys())
    df_data = []

    for strategy in strategies:
        data = analysis_results[strategy]
        row = {
            'Strategy': strategy,
            'Mean Similarity': data['mean_similarity'],
            'Mean Length Difference': data['mean_length_diff'],
            'Mean Robustness': data['mean_robustness'],
            'Mean Score': data['mean_score'],
            'Max Score': data['max_score'],
            'High Score Ratio': data['high_score_ratio'],
            'Medium Score Ratio': data['medium_score_ratio'],
            'Low Score Ratio': data['low_score_ratio'],
            'Score Std Dev': data['std_score']
        }
        df_data.append(row)

    df = pd.DataFrame(df_data)
    return df, analysis_results

def run_multiple_experiments(num_runs=1):
    config = ExperimentConfig()
    all_experiment_results = []

    print(f"Starting {num_runs} experiments...")

    for run_idx in range(num_runs):
        print(f"Running experiment {run_idx + 1}...")

        results = run_single_experiment(config)
        analysis_results = analyze_fitness_values(results, config)
        df, final_analysis = compare_strategies(analysis_results)

        df['Run Count'] = run_idx + 1
        all_experiment_results.append(df)

        print(f"Experiment {run_idx + 1} completed.")

    combined_df = pd.concat(all_experiment_results, ignore_index=True)
    return combined_df

def save_results_to_excel(results_df, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws_raw.append(r)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')

    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_raw.column_dimensions[column_letter].width = adjusted_width

    ws_stats = wb.create_sheet("Statistical Analysis")
    strategies = results_df['Strategy'].unique()
    stats_data = []

    for strategy in strategies:
        strategy_data = results_df[results_df['Strategy'] == strategy]

        stats_row = {
            'Strategy': strategy,
            'Mean Similarity Mean': strategy_data['Mean Similarity'].mean(),
            'Mean Similarity Std': strategy_data['Mean Similarity'].std(),
            'Mean Length Diff Mean': strategy_data['Mean Length Difference'].mean(),
            'Mean Length Diff Std': strategy_data['Mean Length Difference'].std(),
            'Mean Robustness Mean': strategy_data['Mean Robustness'].mean(),
            'Mean Robustness Std': strategy_data['Mean Robustness'].std(),
            'Mean Score Mean': strategy_data['Mean Score'].mean(),
            'Mean Score Std': strategy_data['Mean Score'].std(),
            'Max Score Mean': strategy_data['Max Score'].mean(),
            'Max Score Std': strategy_data['Max Score'].std(),
            'High Score Ratio Mean': strategy_data['High Score Ratio'].mean(),
            'High Score Ratio Std': strategy_data['High Score Ratio'].std()
        }
        stats_data.append(stats_row)

    stats_df = pd.DataFrame(stats_data)

    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_stats.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Results saved to: {output_path}")

def main():
    results_df = run_multiple_experiments(num_runs=20)

    output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"Experiment_Results_{timestamp}.xlsx")

    save_results_to_excel(results_df, output_path)

    print("=" * 60)
    print("All experiments completed!")
    print(f"Completed 20 experiments.")
    print(f"Results saved to: {output_path}")
    print("=" * 60)

    return results_df, output_path

if __name__ == "__main__":
    results, output_file = main()