import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# ==================== 全局范围（威胁分析） ====================
THREAT_COUNT_MIN, THREAT_COUNT_MAX = 1, 120
SEVERITY_LEVEL_MIN, SEVERITY_LEVEL_MAX = 1.0, 10.0
CONFIDENCE_SCORE_MIN, CONFIDENCE_SCORE_MAX = 1.0, 100.0

# 动作扰动步长（severity_level 和 confidence_score 使用浮点步长）
STEP_THREAT = 1
STEP_SEVERITY = 0.1
STEP_CONFIDENCE = 1.0

# === 状态归一化器 ===
class StateNormalizer:
    def __init__(self):
        self.ranges = {
            'threat_count': (THREAT_COUNT_MIN, THREAT_COUNT_MAX),
            'severity_level': (SEVERITY_LEVEL_MIN, SEVERITY_LEVEL_MAX),
            'confidence_score': (CONFIDENCE_SCORE_MIN, CONFIDENCE_SCORE_MAX)
        }

    def normalize(self, state):
        state = np.array(state, dtype=np.float32)
        normalized = np.zeros_like(state)
        normalized[0] = (state[0] - self.ranges['threat_count'][0]) / (self.ranges['threat_count'][1] - self.ranges['threat_count'][0])
        normalized[1] = (state[1] - self.ranges['severity_level'][0]) / (self.ranges['severity_level'][1] - self.ranges['severity_level'][0])
        normalized[2] = (state[2] - self.ranges['confidence_score'][0]) / (self.ranges['confidence_score'][1] - self.ranges['confidence_score'][0])
        return normalized

    def denormalize(self, normalized_state):
        normalized_state = np.array(normalized_state, dtype=np.float32)
        denormalized = np.zeros_like(normalized_state)
        denormalized[0] = normalized_state[0] * (self.ranges['threat_count'][1] - self.ranges['threat_count'][0]) + self.ranges['threat_count'][0]
        denormalized[1] = normalized_state[1] * (self.ranges['severity_level'][1] - self.ranges['severity_level'][0]) + self.ranges['severity_level'][0]
        denormalized[2] = normalized_state[2] * (self.ranges['confidence_score'][1] - self.ranges['confidence_score'][0]) + self.ranges['confidence_score'][0]
        # threat_count 取整，其他保留浮点数
        denormalized[0] = np.clip(np.round(denormalized[0]), self.ranges['threat_count'][0], self.ranges['threat_count'][1]).astype(int)
        denormalized[1] = np.clip(denormalized[1], self.ranges['severity_level'][0], self.ranges['severity_level'][1])
        denormalized[2] = np.clip(denormalized[2], self.ranges['confidence_score'][0], self.ranges['confidence_score'][1])
        return denormalized

normalizer = StateNormalizer()

# === 威胁分析规则函数（直接使用您提供的 execute_threat_analysis_rules） ===
def execute_threat_analysis_rules(a):
    threat_count, severity_level, confidence_score = int(a[0]), float(a[1]), float(a[2])
    triggered = set()
    b = {}
    # 以下为原始规则（与您提供的一致）
    if (threat_count >= 100) != (threat_count >= 110):
        b[0] = 1; triggered.add(1)
    if (threat_count >= 100) != (threat_count == 100):
        b[1] = 2; triggered.add(2)
    if (threat_count >= 100) != (threat_count >= 75):
        b[2] = 3; triggered.add(3)
    if threat_count >= 100:
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 8.0 and confidence_score != 90):
            b[3] = 4; triggered.add(4)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 8.0 and confidence_score >= 95):
            b[4] = 5; triggered.add(5)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level != 8.0 and confidence_score >= 90):
            b[5] = 6; triggered.add(6)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 9.5 and confidence_score >= 90):
            b[6] = 7; triggered.add(7)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 9.3 and confidence_score >= 90):
            b[7] = 8; triggered.add(8)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 8.0 and confidence_score >= 88):
            b[8] = 9; triggered.add(9)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 8.0 and confidence_score >= 95):
            b[9] = 10; triggered.add(10)
        if (severity_level >= 8.0 and confidence_score >= 90) != (severity_level >= 8.3 and confidence_score >= 90):
            b[10] = 11; triggered.add(11)
        if severity_level >= 8.0 and confidence_score >= 90:
            if (confidence_score >= 95) != (confidence_score >= 93):
                b[11] = 12; triggered.add(12)
            if (confidence_score >= 95) != (confidence_score == 95):
                b[12] = 13; triggered.add(13)
            if (confidence_score >= 95) != (confidence_score >= 97):
                b[13] = 14; triggered.add(14)
            if (severity_level >= 9.0) != (severity_level >= 9.1):
                b[14] = 15; triggered.add(15)
            if (severity_level >= 9.0) != (severity_level == 9.0):
                b[15] = 16; triggered.add(16)
            if (severity_level >= 9.0) != (severity_level >= 8.0):
                b[16] = 17; triggered.add(17)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level != 6.0 and confidence_score >= 80):
            b[17] = 18; triggered.add(18)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level == 6.0 and confidence_score >= 80):
            b[18] = 19; triggered.add(19)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score != 80):
            b[19] = 20; triggered.add(20)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score == 80):
            b[20] = 21; triggered.add(21)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score >= 82):
            b[21] = 22; triggered.add(22)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score >= 88):
            b[22] = 23; triggered.add(23)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 7.7 and confidence_score >= 82):
            b[23] = 24; triggered.add(24)
        if (severity_level >= 4.0) != (severity_level != 4.0):
            b[24] = 25; triggered.add(25)
        if (severity_level >= 4.0) != (severity_level >= 7.0):
            b[25] = 26; triggered.add(26)
        if (severity_level >= 4.0) != (severity_level >= 5.0):
            b[26] = 27; triggered.add(27)
    if (threat_count >= 50) != (threat_count >= 75):
        b[27] = 28; triggered.add(28)
    if (threat_count >= 50) != (threat_count >= 59):
        b[28] = 29; triggered.add(29)
    if (threat_count >= 50) != (threat_count >= 55):
        b[29] = 30; triggered.add(30)
    elif threat_count >= 50:
        if (severity_level >= 7.0 and confidence_score >= 85) != (severity_level != 7.0 and confidence_score >= 85):
            b[30] = 31; triggered.add(31)
        if (severity_level >= 7.0 and confidence_score >= 85) != (severity_level == 7.0 and confidence_score >= 85):
            b[31] = 32; triggered.add(32)
        if (severity_level >= 7.0 and confidence_score >= 85) != (severity_level >= 7.0 and confidence_score != 85):
            b[32] = 33; triggered.add(33)
        if (severity_level >= 7.0 and confidence_score >= 85) != (severity_level >= 7.0 and confidence_score == 85):
            b[33] = 34; triggered.add(34)
        if (severity_level >= 7.0 and confidence_score >= 85) != (severity_level >= 7.0 and confidence_score >= 88):
            b[34] = 35; triggered.add(35)
        if severity_level >= 7.0 and confidence_score >= 85:
            if (confidence_score >= 95) != (confidence_score >= 85):
                b[35] = 36; triggered.add(36)
            if (confidence_score >= 95) != (confidence_score == 95):
                b[36] = 37; triggered.add(37)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level >= 5.0 and confidence_score != 70):
            b[37] = 38; triggered.add(38)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level >= 5.0 and confidence_score == 70):
            b[38] = 39; triggered.add(39)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level >= 5.0 and confidence_score >= 75):
            b[39] = 40; triggered.add(40)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level != 5.0 and confidence_score >= 70):
            b[40] = 41; triggered.add(41)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level == 5.0 and confidence_score >= 70):
            b[41] = 42; triggered.add(42)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level >= 5.0 and confidence_score >= 74):
            b[42] = 43; triggered.add(43)
        if (severity_level >= 5.0 and confidence_score >= 70) != (severity_level >= 6.2 and confidence_score >= 70):
            b[43] = 44; triggered.add(44)
    if (threat_count >= 20) != (threat_count != 20):
        b[44] = 45; triggered.add(45)
    if (threat_count >= 20) != (threat_count >= 30):
        b[45] = 46; triggered.add(46)
    if (threat_count >= 20) != (threat_count >= 25):
        b[46] = 47; triggered.add(47)
    elif threat_count >= 20:
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score != 80):
            b[47] = 48; triggered.add(48)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score == 80):
            b[48] = 49; triggered.add(49)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level >= 6.0 and confidence_score >= 87):
            b[49] = 50; triggered.add(50)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level != 6.0 and confidence_score >= 80):
            b[50] = 51; triggered.add(51)
        if (severity_level >= 6.0 and confidence_score >= 80) != (severity_level == 6.0 and confidence_score >= 80):
            b[51] = 52; triggered.add(52)
        if (severity_level >= 4.0) != (severity_level != 4.0):
            b[52] = 53; triggered.add(53)
        if (severity_level >= 4.0) != (severity_level >= 2.5):
            b[53] = 54; triggered.add(54)
    if (threat_count >= 5) != (threat_count != 5):
        b[54] = 55; triggered.add(55)
    if (threat_count >= 5) != (threat_count >= 7.2):
        b[55] = 56; triggered.add(56)
    elif threat_count >= 5:
        if (severity_level >= 7.0) != (severity_level >= 8.2):
            b[56] = 57; triggered.add(57)
        if (severity_level >= 7.0) != (severity_level == 7.0):
            b[57] = 58; triggered.add(58)
        if (severity_level >= 7.0) != (severity_level >= 7.5):
            b[58] = 59; triggered.add(59)
    if (severity_level >= 9.0) != (severity_level >= 9.4):
        b[59] = 60; triggered.add(60)
    if (severity_level >= 9.0) != (severity_level == 9.0):
        b[60] = 61; triggered.add(61)
    if severity_level >= 9.0:
        if (threat_count >= 50 and confidence_score >= 90) != (threat_count >= 50 and confidence_score >= 93):
            b[61] = 62; triggered.add(62)
        if (threat_count >= 50 and confidence_score >= 90) != (threat_count >= 50 and confidence_score == 90):
            b[62] = 63; triggered.add(63)
        if (threat_count >= 50 and confidence_score >= 90) != (threat_count >= 55 and confidence_score >= 93):
            b[63] = 64; triggered.add(64)
        if (threat_count >= 50 and confidence_score >= 90) != (threat_count >= 66 and confidence_score >= 90):
            b[64] = 65; triggered.add(65)
        if (threat_count >= 50 and confidence_score >= 90) != (threat_count >= 50 and confidence_score == 90):
            b[65] = 66; triggered.add(66)
        if (threat_count >= 20 or confidence_score >= 85) != (threat_count != 20 or confidence_score >= 85):
            b[66] = 67; triggered.add(67)
        if (threat_count >= 20 or confidence_score >= 85) != (threat_count >= 50 or confidence_score >= 85):
            b[67] = 68; triggered.add(68)
        if (threat_count >= 20 or confidence_score >= 85) != (threat_count >= 35 or confidence_score >= 85):
            b[68] = 69; triggered.add(69)
        if (threat_count >= 20 or confidence_score >= 85) != (threat_count >= 20 or confidence_score != 85):
            b[69] = 70; triggered.add(70)
        if (threat_count >= 20 or confidence_score >= 85) != (threat_count >= 20 or confidence_score == 85):
            b[70] = 71; triggered.add(71)
    if (severity_level >= 7.0) != (severity_level >= 5.5):
        b[71] = 72; triggered.add(72)
    if (severity_level >= 7.0) != (severity_level == 7.0):
        b[72] = 73; triggered.add(73)
    elif severity_level >= 7.0:
        if (confidence_score >= 85) != (confidence_score >= 87):
            b[73] = 74; triggered.add(74)
        if (confidence_score >= 85) != (confidence_score == 85):
            b[74] = 75; triggered.add(75)
    if (severity_level >= 5.0) != (severity_level != 5.0):
        b[75] = 76; triggered.add(76)
    if (severity_level >= 5.0) != (severity_level >= 6.0):
        b[76] = 77; triggered.add(77)
    if (confidence_score >= 95) != (confidence_score >= 97):
        b[77] = 78; triggered.add(78)
    if (confidence_score >= 95) != (confidence_score >= 89):
        b[78] = 79; triggered.add(79)
    if (confidence_score >= 95) != (confidence_score == 95):
        b[79] = 80; triggered.add(80)
    if confidence_score >= 95:
        if (severity_level >= 8.0 and threat_count >= 30) != (severity_level >= 8.0 and threat_count != 30):
            b[80] = 81; triggered.add(81)
        if (severity_level >= 8.0 and threat_count >= 30) != (severity_level >= 8.0 and threat_count == 30):
            b[81] = 82; triggered.add(82)
        if (severity_level >= 8.0 and threat_count >= 30) != (severity_level >= 8.0 and threat_count >= 31):
            b[82] = 83; triggered.add(83)
        if (severity_level >= 8.0 and threat_count >= 30) != (severity_level >= 8.0 and threat_count >= 45):
            b[83] = 84; triggered.add(84)
        if (severity_level >= 8.0 and threat_count >= 30) != (severity_level == 8.0 and threat_count >= 30):
            b[84] = 85; triggered.add(85)
        if (severity_level >= 6.0) != (severity_level >= 6.6):
            b[85] = 86; triggered.add(86)
        if (severity_level >= 6.0) != (severity_level >= 7.3):
            b[86] = 87; triggered.add(87)
    if (confidence_score >= 85) != (confidence_score >= 78):
        b[87] = 88; triggered.add(88)
    if (confidence_score >= 85) != (confidence_score >= 87):
        b[88] = 89; triggered.add(89)
    elif confidence_score >= 85:
        if (severity_level >= 7.0) != (severity_level >= 8.7):
            b[89] = 90; triggered.add(90)
        if (severity_level >= 7.0) != (severity_level == 7.0):
            b[90] = 91; triggered.add(91)
        if (severity_level >= 7.0) != (severity_level >= 7.2):
            b[91] = 92; triggered.add(92)
        if (severity_level >= 5.0) != (severity_level >= 6.3):
            b[92] = 93; triggered.add(93)
        if (severity_level >= 5.0) != (severity_level >= 3.7):
            b[93] = 94; triggered.add(94)
    if (confidence_score >= 70) != (confidence_score >= 77):
        b[94] = 95; triggered.add(95)
    if (confidence_score >= 70) != (confidence_score >= 86):
        b[95] = 96; triggered.add(96)
    return triggered

# 设置别名，统一使用 execute_Tr
execute_Tr = execute_threat_analysis_rules

# === 目标路径组（与您提供的一致） ===
targetPaths = [
    # A1
    {6, 7, 8, 11, 13, 17, 19, 21, 28, 29, 30, 32, 34, 37, 39, 42, 49, 52, 57, 58, 63, 64, 65, 66, 73, 75, 80, 82, 90, 91},
    # A2
    {5, 7, 8, 10, 15, 19, 21, 28, 29, 32, 34, 36, 39, 42, 49, 52, 58, 60, 62, 63, 64, 65, 66, 73, 75, 79, 82, 85, 91},
    # A3
    {6, 7, 8, 11, 14, 17, 19, 21, 28, 29, 30, 32, 34, 39, 42, 49, 52, 57, 58, 63, 64, 65, 66, 73, 75, 78, 82, 90, 91},
    # A4
    {4, 5, 7, 8, 10, 15, 19, 21, 28, 29, 30, 32, 34, 36, 39, 42, 49, 52, 58, 60, 62, 64, 65, 73, 75, 79, 82, 85, 91},
    # A5
    {5, 6, 7, 8, 10, 11, 12, 17, 19, 21, 28, 32, 34, 36, 39, 42, 49, 52, 57, 58, 63, 65, 66, 73, 75, 79, 82, 90, 91},
    # A6
    {5, 10, 16, 19, 21, 28, 29, 30, 32, 34, 36, 39, 42, 49, 52, 58, 61, 62, 63, 64, 65, 66, 73, 75, 79, 82, 85, 91},
    # A7
    {1, 2, 6, 7, 8, 11, 13, 14, 17, 19, 21, 32, 34, 37, 39, 42, 49, 52, 57, 58, 63, 66, 73, 75, 78, 80, 82, 90, 91},
    # A8
    {5, 6, 7, 8, 10, 11, 12, 17, 19, 21, 32, 34, 36, 39, 42, 45, 49, 52, 55, 57, 58, 71, 73, 75, 79, 81, 90, 91},
    # A9
    {3, 6, 7, 8, 11, 13, 14, 17, 19, 21, 32, 34, 37, 39, 42, 49, 52, 57, 58, 63, 66, 73, 75, 78, 80, 82, 90, 91},
    # A10
    {4, 5, 6, 7, 8, 10, 11, 17, 19, 21, 32, 34, 36, 39, 42, 45, 46, 47, 49, 52, 57, 58, 73, 75, 79, 81, 90, 91},
    # A11
    {6, 7, 8, 11, 13, 17, 19, 21, 32, 34, 37, 39, 42, 45, 49, 52, 55, 56, 57, 58, 71, 73, 75, 80, 81, 90, 91},
    # A12
    {4, 15, 19, 21, 23, 28, 29, 30, 32, 34, 35, 36, 39, 42, 49, 50, 52, 58, 60, 73, 74, 75, 82, 85, 89, 91},
    # A13
    {4, 17, 19, 21, 23, 32, 33, 35, 36, 39, 42, 45, 46, 47, 49, 50, 52, 57, 58, 73, 74, 81, 89, 90, 91, 96},
    # A14
    {4, 17, 19, 21, 23, 32, 33, 35, 36, 39, 42, 45, 49, 50, 52, 56, 57, 58, 70, 73, 74, 81, 89, 90, 91, 96},
    # A15
    {4, 17, 19, 21, 23, 33, 39, 42, 45, 46, 47, 49, 50, 52, 57, 58, 67, 68, 69, 73, 81, 88, 90, 91, 96},
    # A16
    {4, 17, 19, 21, 22, 23, 24, 33, 39, 42, 45, 49, 50, 52, 55, 57, 58, 67, 70, 73, 81, 88, 90, 91, 96},
    # A17
    {6, 13, 14, 18, 21, 24, 26, 28, 31, 37, 39, 42, 44, 49, 51, 63, 65, 66, 72, 75, 78, 80, 86, 87, 93},
    # A18
    {6, 13, 14, 19, 21, 24, 28, 31, 34, 37, 39, 42, 49, 52, 57, 59, 63, 65, 66, 75, 78, 80, 87, 90, 92},
    # A19
    {5, 6, 7, 8, 10, 11, 17, 19, 21, 32, 34, 36, 39, 42, 49, 52, 57, 58, 73, 75, 79, 82, 84, 90, 91},
    # A20
    {4, 16, 19, 20, 22, 23, 24, 28, 29, 30, 33, 39, 42, 48, 50, 52, 58, 61, 73, 82, 85, 88, 91, 96},
    # A21
    {6, 13, 14, 18, 26, 28, 29, 30, 31, 37, 39, 41, 44, 51, 63, 64, 65, 66, 75, 76, 77, 78, 80, 93},
    # A22
    {6, 7, 8, 11, 14, 17, 19, 21, 32, 34, 39, 42, 49, 52, 57, 58, 73, 75, 78, 81, 83, 84, 90, 91},
    # A23
    {4, 9, 17, 19, 21, 32, 34, 36, 39, 42, 45, 46, 47, 49, 52, 57, 58, 73, 75, 79, 81, 90, 91},
    # A24
    {4, 17, 20, 33, 38, 40, 42, 43, 45, 48, 55, 56, 57, 58, 67, 70, 73, 81, 90, 91, 95, 96},
    # A25
    {6, 18, 25, 26, 27, 28, 29, 30, 31, 36, 41, 51, 53, 62, 63, 64, 65, 66, 75, 76, 79, 94},
    # A26
    {6, 12, 18, 25, 28, 29, 30, 31, 36, 41, 51, 53, 54, 63, 64, 65, 66, 75, 76, 79}
]
target_paths = [set(path) for path in targetPaths]

def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0

def compute_path_similarity_matrix(paths):
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix

def group_paths_by_similarity(paths):
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.mean(avg_sim_scores)
    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]
    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)
    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group

def compute_robustness(state, path):
    base = execute_Tr(*state)
    if not base:
        return 0.0
    rob, neighbors = 0.0, 0
    for dt in [-1, 0, 1]:
        for ds in [-1, 0, 1]:
            for dc in [-1, 0, 1]:
                if dt == ds == dc == 0:
                    continue
                neighbor = np.array([
                    np.clip(state[0] + dt * STEP_THREAT, THREAT_COUNT_MIN, THREAT_COUNT_MAX),
                    np.clip(state[1] + ds * STEP_SEVERITY, SEVERITY_LEVEL_MIN, SEVERITY_LEVEL_MAX),
                    np.clip(state[2] + dc * STEP_CONFIDENCE, CONFIDENCE_SCORE_MIN, CONFIDENCE_SCORE_MAX)
                ])
                # 保留两位小数
                neighbor[1] = round(neighbor[1], 2)
                neighbor[2] = round(neighbor[2], 2)
                n_trig = execute_Tr(*neighbor)
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0

def compute_q_value_score(state, similar_model):
    if similar_model is None:
        return 0.0
    try:
        normalized_state = normalizer.normalize(state)
        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = similar_model(state_tensor)
            max_q_value = torch.max(q_values).item()
            normalized_q = max_q_value / 20.0
            normalized_q = max(0.0, min(1.0, normalized_q))
            return 1.0 - normalized_q
    except:
        return 0.0

def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward

def generate_samples_for_similar_paths(similar_group, num_candidates=2000, top_k=200, run_id=1):
    SIMILAR_WEIGHTS = [0.55, 0.39, 0.06]
    def save_samples(path_id, samples, base_dir, group_type="similar"):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_{group_type}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"{group_type.title()} Group Path {path_id} - Run {run_id}\n")
            f.write("threat_count severity_level confidence_score\tScore\tSimilarity\tLengthDiff\tRobustness\n")
            for s in samples:
                tc, sl, cs = s['state']
                f.write(f"{tc} {sl:.2f} {cs:.2f}\t{s['score']:.4f}\t{s['similarity']:.4f}\t{s['length_diff']:.4f}\t{s['robustness']:.4f}\n")
    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_grouped"
    for path_idx in similar_group:
        path = target_paths[path_idx]
        path_id = path_idx + 1
        candidate_samples = []
        attempts = 0
        while len(candidate_samples) < num_candidates and attempts < num_candidates * 10:
            attempts += 1
            tc = random.randint(THREAT_COUNT_MIN, THREAT_COUNT_MAX)
            sl = round(random.uniform(SEVERITY_LEVEL_MIN, SEVERITY_LEVEL_MAX), 2)
            cs = round(random.uniform(CONFIDENCE_SCORE_MIN, CONFIDENCE_SCORE_MAX), 2)
            state = (tc, sl, cs)
            triggered = execute_Tr(*state)
            if not triggered:
                continue
            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            candidate_samples.append({
                'state': state,
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            })
        if candidate_samples:
            for sample in candidate_samples:
                score = (SIMILAR_WEIGHTS[0] * sample['similarity'] +
                         SIMILAR_WEIGHTS[1] * sample['length_diff'] +
                         SIMILAR_WEIGHTS[2] * sample['robustness'])
                sample['score'] = score
            candidate_samples.sort(key=lambda x: x['score'], reverse=True)
            selected_samples = candidate_samples[:top_k]
            save_samples(path_id=path_id, samples=selected_samples, base_dir=base_dir, group_type="similar")

def generate_samples_for_isolated_paths(isolated_group, similar_model, num_candidates=2000, top_k=200, run_id=1):
    ISOLATED_WEIGHTS = [0.18, 0.21, 0.32, 0.29]
    def save_samples(path_id, samples, base_dir, group_type="isolated"):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_{group_type}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"{group_type.title()} Group Path {path_id} - Run {run_id}\n")
            f.write("threat_count severity_level confidence_score\tScore\tSimilarity\tLengthDiff\tRobustness\tQValueScore\n")
            for s in samples:
                tc, sl, cs = s['state']
                f.write(f"{tc} {sl:.2f} {cs:.2f}\t{s['score']:.4f}\t{s['similarity']:.4f}\t{s['length_diff']:.4f}\t{s['robustness']:.4f}\t{s['q_value_score']:.4f}\n")
    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_grouped"
    for path_idx in isolated_group:
        path = target_paths[path_idx]
        path_id = path_idx + 1
        candidate_samples = []
        attempts = 0
        while len(candidate_samples) < num_candidates and attempts < num_candidates * 10:
            attempts += 1
            tc = random.randint(THREAT_COUNT_MIN, THREAT_COUNT_MAX)
            sl = round(random.uniform(SEVERITY_LEVEL_MIN, SEVERITY_LEVEL_MAX), 2)
            cs = round(random.uniform(CONFIDENCE_SCORE_MIN, CONFIDENCE_SCORE_MAX), 2)
            state = (tc, sl, cs)
            triggered = execute_Tr(*state)
            if not triggered:
                continue
            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            q_score = compute_q_value_score(state, similar_model)
            candidate_samples.append({
                'state': state,
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'q_value_score': q_score,
                'triggered': triggered
            })
        if candidate_samples:
            for sample in candidate_samples:
                score = (ISOLATED_WEIGHTS[0] * sample['similarity'] +
                         ISOLATED_WEIGHTS[1] * sample['length_diff'] +
                         ISOLATED_WEIGHTS[2] * sample['robustness'] +
                         ISOLATED_WEIGHTS[3] * sample['q_value_score'])
                sample['score'] = score
            candidate_samples.sort(key=lambda x: x['score'], reverse=True)
            selected_samples = candidate_samples[:top_k]
            save_samples(path_id=path_id, samples=selected_samples, base_dir=base_dir, group_type="isolated")

class GroupExperienceReplay:
    def __init__(self, capacity=20000):
        self.capacity = capacity
        self.buffer = deque(maxlen=self.capacity)
        self.priorities = deque(maxlen=self.capacity)
        self.sampled_indices = set()
    def append(self, experience):
        self.buffer.append(experience)
        self.priorities.append(experience[-1])
    def sample(self, batch_size, alpha=0.6):
        priorities = np.array(self.priorities) ** alpha
        probabilities = priorities / np.sum(priorities)
        batch_indices = np.random.choice(len(self.buffer), batch_size, p=probabilities)
        batch = [self.buffer[idx] for idx in batch_indices]
        return batch, batch_indices, probabilities[batch_indices]
    def update_priorities(self, batch_indices, td_errors):
        for idx, td_error in zip(batch_indices, td_errors):
            if idx < len(self.priorities):
                self.priorities[idx] = max(td_error, 1e-6)
    def __len__(self):
        return len(self.buffer)
    def get_high_reward_samples(self, target_path, num_samples=20):
        if len(self.buffer) == 0:
            return []
        samples_with_recalculated_scores = []
        for idx, experience in enumerate(self.buffer):
            if idx in self.sampled_indices:
                continue
            normalized_state_tensor = experience[0]
            normalized_state = normalized_state_tensor.cpu().numpy().flatten()
            state_tuple = tuple(normalizer.denormalize(normalized_state))
            tc, sl, cs = state_tuple
            triggered = execute_Tr(tc, sl, cs)
            new_reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_recalculated_scores.append((idx, state_tuple, new_reward, sim, triggered))
        samples_with_recalculated_scores.sort(key=lambda x: x[2], reverse=True)
        selected = samples_with_recalculated_scores[:num_samples]
        for item in selected:
            self.sampled_indices.add(item[0])
        return [(s[1], s[2], s[3], s[4]) for s in selected]
    def reset_sampled_indices(self):
        self.sampled_indices.clear()

def load_path_data(file_path):
    path_data = []
    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines[2:]:
            parts = line.strip().split("\t")
            state = parts[0].split()
            tc = int(state[0])
            sl = float(state[1])
            cs = float(state[2])
            path_data.append((tc, sl, cs))
    return path_data

class DQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, 128)
        self.fc2 = nn.Linear(128, 64)
        self.fc3 = nn.Linear(64, action_dim)
    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)

class DQNAgentWithPER:
    def __init__(self, state_dim, action_dim, replay_buffer, gamma=0.99, epsilon=1.0, epsilon_decay=0.995,
                 epsilon_min=0.1, learning_rate=0.001, alpha=0.6, beta=0.4):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer
        self.alpha = alpha
        self.beta = beta
        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())
    def decode_action(self, action_idx):
        delta_values = [1, -1]
        dim = action_idx // 2
        delta_idx = action_idx % 2
        delta = delta_values[delta_idx]
        if dim == 0:
            return (delta, 0, 0)
        elif dim == 1:
            return (0, delta, 0)
        elif dim == 2:
            return (0, 0, delta)
    def act(self, normalized_state):
        if random.random() < self.epsilon:
            return random.randrange(self.action_dim)
        state = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state)
        return torch.argmax(q_values, dim=1).item()
    def store_transition(self, normalized_state, action, reward, normalized_next_state, done):
        state = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        next_state = torch.tensor(normalized_next_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state)
            next_q_values = self.target_model(next_state)
            max_next_q_values = next_q_values.max(1)[0]
            target_q_values = reward + (self.gamma * max_next_q_values * (1 - done))
            td_error = torch.abs(q_values[0][action] - target_q_values).item()
        self.replay_buffer.append((state, action, reward, next_state, done, td_error))
        return td_error
    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return
        batch, batch_indices, probabilities = self.replay_buffer.sample(batch_size, alpha=self.alpha)
        states, actions, rewards, next_states, dones, _ = zip(*batch)
        weights = (len(self.replay_buffer) * probabilities) ** (-self.beta)
        weights = weights / weights.max()
        weights = torch.tensor(weights, dtype=torch.float32).to(device)
        states = torch.tensor(np.array([s.cpu().numpy().flatten() for s in states]), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array([ns.cpu().numpy().flatten() for ns in next_states]), dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)
        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))
        td_errors = current_q_values - target_q_values
        weighted_loss = (td_errors.pow(2) * weights).mean()
        self.optimizer.zero_grad()
        weighted_loss.backward()
        self.optimizer.step()
        new_priorities = torch.abs(td_errors).detach().cpu().numpy()
        self.replay_buffer.update_priorities(batch_indices, new_priorities)
        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay
    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())

def train_group(group_paths, path_documents, replay_buffer, batch_size=32, group_name="", pretrained_model=None):
    state_dim = 3
    action_dim = 6
    agent = DQNAgentWithPER(state_dim, action_dim, replay_buffer)
    if pretrained_model is not None:
        print(f"  {group_name}: 加载预训练模型...")
        agent.model.load_state_dict(pretrained_model.state_dict())
        agent.target_model.load_state_dict(pretrained_model.state_dict())
        print(f"  {group_name}: 加载完成")
    path_rewards = {}
    print(f"开始训练 {group_name}，包含路径: {[idx + 1 for idx in group_paths]}")
    start_time = time.time()
    BATCH_SIZE = 50
    N_SAMPLES = 200
    N_STEPS = 3
    N_ROUNDS = 5
    N_BATCHES = 4
    replay_count = 0
    for path_idx in group_paths:
        file_path = os.path.join(path_documents, f"path{path_idx + 1}_{'similar' if group_name == '' else 'isolated'}.txt")
        if not os.path.exists(file_path):
            print(f"    警告: 路径 {path_idx + 1} 的样本文件不存在，跳过")
            continue
        path_data = load_path_data(file_path)
        target_path = target_paths[path_idx]
        if path_idx not in path_rewards:
            path_rewards[path_idx] = 0
        print(f"\n  开始训练路径 {path_idx + 1}，共 {N_ROUNDS} 轮")
        for round_idx in range(N_ROUNDS):
            print(f"    路径 {path_idx + 1} - 第 {round_idx + 1}/{N_ROUNDS} 轮")
            for batch_idx in range(N_BATCHES):
                batch_start = batch_idx * BATCH_SIZE
                batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)
                if batch_start >= len(path_data):
                    print(f"       第 {batch_idx + 1} 批: 样本不足，跳过")
                    break
                print(f"       第 {batch_idx + 1}/{N_BATCHES} 批 (样本 {batch_start}-{batch_end})")
                for sample_idx in range(batch_start, batch_end):
                    if sample_idx >= len(path_data):
                        break
                    state = path_data[sample_idx]
                    prev_state = None
                    prev_triggered = None
                    for step in range(N_STEPS):
                        normalized_state = normalizer.normalize(state)
                        legal_actions = []
                        for a in range(agent.action_dim):
                            dt, ds, dc = agent.decode_action(a)
                            cand_next = tuple(np.clip(np.array(state) + np.array([dt, ds, dc]),
                                                      [THREAT_COUNT_MIN, SEVERITY_LEVEL_MIN, CONFIDENCE_SCORE_MIN],
                                                      [THREAT_COUNT_MAX, SEVERITY_LEVEL_MAX, CONFIDENCE_SCORE_MAX]))
                            legal_actions.append(a)
                        if not legal_actions:
                            break
                        if random.random() < agent.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]
                        dt, ds, dc = agent.decode_action(action)
                        next_state = tuple(np.clip(np.array(state) + np.array([dt, ds, dc]),
                                                   [THREAT_COUNT_MIN, SEVERITY_LEVEL_MIN, CONFIDENCE_SCORE_MIN],
                                                   [THREAT_COUNT_MAX, SEVERITY_LEVEL_MAX, CONFIDENCE_SCORE_MAX]))
                        next_state = (int(next_state[0]), round(next_state[1], 2), round(next_state[2], 2))
                        normalized_next_state = normalizer.normalize(next_state)
                        tc, sl, cs = next_state
                        triggered = execute_Tr(tc, sl, cs)
                        reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                        done = (step == N_STEPS - 1)
                        agent.store_transition(normalized_state, action, reward, normalized_next_state, done)
                        prev_state = state
                        prev_triggered = triggered
                        state = next_state
                        path_rewards[path_idx] += reward
                if len(agent.replay_buffer) >= batch_size:
                    agent.train(batch_size)
                    replay_count += 1
                    if replay_count % 2 == 0:
                        agent.update_target_model()
            print(f"      路径 {path_idx + 1} - 第 {round_idx + 1} 轮完成")
        print(f"  路径 {path_idx + 1} 的 {N_ROUNDS} 轮全部完成")
    training_time = time.time() - start_time
    print(f"\n{group_name} 训练完成！")
    print(f"  完成路径: {N_ROUNDS} 轮 (每轮 {N_BATCHES} 批)")
    print(f"  总回放更新次数: {replay_count}")
    print(f"  训练耗时: {training_time:.2f} 秒")
    print(f"  经验池大小: {len(replay_buffer)}")
    return agent, path_rewards, training_time

def generate_and_train_grouped_paths_staged(path_documents, similar_group, isolated_group, batch_size=32, run_id=1):
    print(f"\n=== 第 {run_id}/20 次运行 (3分钟版本) ===")
    similar_group_paths = [idx + 1 for idx in similar_group]
    isolated_group_paths = [idx + 1 for idx in isolated_group]
    print(f"相似路径组: {similar_group_paths}")
    print(f"孤立路径组: {isolated_group_paths}")
    total_start_time = time.time()
    print(f"\n[1] 生成相似路径样本...")
    generate_samples_for_similar_paths(similar_group, num_candidates=2000, top_k=200, run_id=run_id)
    print(f"\n[2] 训练相似路径组 (共 {len(similar_group)} 条路径, 每路径5轮)...")
    similar_replay_buffer = GroupExperienceReplay(capacity=20000)
    similar_agent, similar_path_rewards, similar_training_time = train_group(
        similar_group, path_documents, similar_replay_buffer, batch_size,
        group_name="相似路径组", pretrained_model=None
    )
    print(f"\n[3] 生成孤立路径样本...")
    generate_samples_for_isolated_paths(isolated_group, similar_agent.model,
                                        num_candidates=2000, top_k=200, run_id=run_id)
    print(f"\n[4] 训练孤立路径组 (共 {len(isolated_group)} 条路径, 每路径5轮)...")
    isolated_replay_buffer = GroupExperienceReplay(capacity=20000)
    isolated_agent, isolated_path_rewards, isolated_training_time = train_group(
        isolated_group, path_documents, isolated_replay_buffer, batch_size,
        group_name="孤立路径组", pretrained_model=similar_agent.model
    )
    total_path_rewards = {**similar_path_rewards, **isolated_path_rewards}
    total_cumulative_reward = sum(total_path_rewards.values())
    total_training_time = time.time() - total_start_time
    print(f"\n=== 第 {run_id}/20 次运行完成，总耗时: {total_training_time:.2f} 秒 ===")
    print(f"  相似组训练: {similar_training_time:.2f} 秒")
    print(f"  孤立组训练: {isolated_training_time:.2f} 秒")
    print(f"  经验池大小 - 相似: {len(similar_replay_buffer)}, 孤立: {len(isolated_replay_buffer)}")
    return similar_agent, isolated_agent, similar_replay_buffer, isolated_replay_buffer, \
           total_cumulative_reward, total_path_rewards, total_training_time

def create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    similar_group_paths = [idx + 1 for idx in similar_group]
    isolated_group_paths = [idx + 1 for idx in isolated_group]
    wb = Workbook()
    thin_border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
                         top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    header_color = "4472C4"
    similar_group_color = "E2EFDA"
    isolated_group_color = "FCE4D6"
    stats_color = "FFF2CC"
    ws_paths = wb.active
    ws_paths.title = "路径相似度"
    path_headers = ['Path ID', '分组'] + [f'Run {i}' for i in range(1, 21)] + ['平均相似度', '最大相似度', '最小相似度', '标准差']
    for col, header in enumerate(path_headers, 1):
        cell = ws_paths.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws_paths.row_dimensions[1].height = 30
    for path_id in range(1, len(target_paths) + 1):
        row = path_id + 1
        if path_id in similar_group_paths:
            group_type = "高相关路径组"
            row_color = similar_group_color
        elif path_id in isolated_group_paths:
            group_type = "低相关路径组"
            row_color = isolated_group_color
        else:
            group_type = "未分组"
            row_color = "FFFFFF"
        cell = ws_paths.cell(row=row, column=1, value=f"Path {path_id}")
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        cell.border = thin_border
        cell = ws_paths.cell(row=row, column=2, value=group_type)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        cell.border = thin_border
        path_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            sim = run_data['path_similarities'].get(path_id, {}).get('avg_similarity', 0.0)
            path_similarities.append(sim)
            cell = ws_paths.cell(row=row, column=3 + run_idx, value=round(sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        stats_values = [
            round(np.mean(path_similarities), 4),
            round(np.max(path_similarities), 4),
            round(np.min(path_similarities), 4),
            round(np.std(path_similarities), 4)
        ]
        for i, value in enumerate(stats_values):
            cell = ws_paths.cell(row=row, column=23 + i, value=value)
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
            cell.font = Font(bold=True, size=10)
            cell.border = thin_border
    ws_paths.column_dimensions['A'].width = 13
    ws_paths.column_dimensions['B'].width = 16
    for col in range(3, 23):
        ws_paths.column_dimensions[get_column_letter(col)].width = 10
    for col in range(23, 27):
        ws_paths.column_dimensions[get_column_letter(col)].width = 13
    ws_groups = wb.create_sheet("分组统计")
    group_headers = ['组名', '包含路径'] + [f'Run {i}' for i in range(1, 21)] + ['平均相似度', '标准差']
    for col, header in enumerate(group_headers, 1):
        cell = ws_groups.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws_groups.row_dimensions[1].height = 30
    row = 2
    cell = ws_groups.cell(row=row, column=1, value="高相关路径组")
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    cell.border = thin_border
    cell = ws_groups.cell(row=row, column=2, value=','.join(map(str, similar_group_paths)))
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    cell.border = thin_border
    group_similarities = []
    for run_idx, run_data in enumerate(all_runs_data):
        group_sim = np.mean([run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0) for p in similar_group_paths])
        group_similarities.append(group_sim)
        cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(group_sim, 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    cell = ws_groups.cell(row=row, column=23, value=round(np.mean(group_similarities), 4))
    cell.number_format = '0.0000'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border
    cell = ws_groups.cell(row=row, column=24, value=round(np.std(group_similarities), 4))
    cell.number_format = '0.0000'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border
    row += 1
    if isolated_group_paths:
        cell = ws_groups.cell(row=row, column=1, value="低相关路径组")
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        cell.border = thin_border
        cell = ws_groups.cell(row=row, column=2, value=','.join(map(str, isolated_group_paths)))
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        cell.border = thin_border
        isolated_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            iso_sim = np.mean([run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0) for p in isolated_group_paths])
            isolated_similarities.append(iso_sim)
            cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(iso_sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        cell = ws_groups.cell(row=row, column=23, value=round(np.mean(isolated_similarities), 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border
        cell = ws_groups.cell(row=row, column=24, value=round(np.std(isolated_similarities), 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border
    ws_groups.column_dimensions['A'].width = 16
    ws_groups.column_dimensions['B'].width = 22
    for col in range(3, 23):
        ws_groups.column_dimensions[get_column_letter(col)].width = 10
    ws_groups.column_dimensions[get_column_letter(23)].width = 14
    ws_groups.column_dimensions[get_column_letter(24)].width = 12
    ws_samples = wb.create_sheet("详细样本数据")
    sample_headers = ['Run', 'Path ID', 'Sample ID', 'Threat_Count', 'Severity_Level', 'Confidence_Score', 'Similarity', '触发规则集']
    for col, header in enumerate(sample_headers, 1):
        cell = ws_samples.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws_samples.row_dimensions[1].height = 30
    sample_row = 2
    for run_idx, run_data in enumerate(all_runs_data, 1):
        for path_id in range(1, len(target_paths)+1):
            samples = run_data['path_samples'].get(path_id, [])
            if path_id in similar_group_paths:
                path_color = similar_group_color
            elif path_id in isolated_group_paths:
                path_color = isolated_group_color
            else:
                path_color = "FFFFFF"
            for sample_idx, (state_tuple, reward, sim, triggered) in enumerate(samples, 1):
                tc, sl, cs = state_tuple
                triggered_str = ','.join(map(str, sorted(triggered)))
                cell = ws_samples.cell(row=sample_row, column=1, value=f"Run {run_idx}")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=path_color, end_color=path_color, fill_type="solid")
                cell.border = thin_border
                cell = ws_samples.cell(row=sample_row, column=2, value=f"Path {path_id}")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=path_color, end_color=path_color, fill_type="solid")
                cell.border = thin_border
                cell = ws_samples.cell(row=sample_row, column=3, value=sample_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                for col_offset, value in enumerate([tc, sl, cs]):
                    cell = ws_samples.cell(row=sample_row, column=4 + col_offset, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                cell = ws_samples.cell(row=sample_row, column=7, value=round(sim, 4))
                cell.number_format = '0.0000'
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                cell = ws_samples.cell(row=sample_row, column=8, value=f"{{{triggered_str}}}")
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border
                sample_row += 1
    sample_widths = [13, 13, 11, 14, 16, 16, 12, 45]
    for i, width in enumerate(sample_widths, 1):
        ws_samples.column_dimensions[get_column_letter(i)].width = width
    output_path = os.path.join(output_dir, "20次运行_威胁分析_3分钟版本.xlsx")
    wb.save(output_path)
    print(f"\n综合Excel报告已生成: {output_path}")

def run_20_times_training():
    model_path_base = r"D:\Experiment\CNN\DQNNEW\saved_models_3min_version"
    path_documents = r"D:\Experiment\CNN\DQNNEW\path_samples_grouped"
    output_dir = r"D:\Experiment\CNN\ComparisonExperiment2\excel_reports_3min_version"
    os.makedirs(model_path_base, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)
    similar_group, isolated_group = group_paths_by_similarity(target_paths)
    similar_group_display = [idx + 1 for idx in similar_group]
    isolated_group_display = [idx + 1 for idx in isolated_group]
    print("=" * 60)
    print("20次运行 - 3分钟版本 (威胁分析)")
    print("=" * 60)
    print("训练规模配置:")
    print("   每路径: 5轮")
    print("   每轮: 4批")
    print("   每批: 50样本")
    print("   每样本: 3步")
    print("   样本生成: 2000候选 -> 200最终样本")
    print("   模型保存: 仅保存模型参数(优化版)")
    print("=" * 60)
    print(f"\n自动分组结果:")
    print(f"相似路径组: {similar_group_display}")
    print(f"孤立路径组: {isolated_group_display}")
    print("\n" + "=" * 60)
    all_runs_data = []
    total_start_time = time.time()
    for run_id in range(1, 21):
        print(f"\n{'=' * 60}")
        print(f"开始第 {run_id}/20 次运行")
        print(f"{'=' * 60}")
        similar_agent, isolated_agent, similar_buffer, isolated_buffer, total_cumulative_reward, path_rewards, training_time = \
            generate_and_train_grouped_paths_staged(path_documents, similar_group, isolated_group, batch_size=32, run_id=run_id)
        similar_model_path = os.path.join(model_path_base, f"similar_group_model_run_{run_id}.pth")
        isolated_model_path = os.path.join(model_path_base, f"isolated_group_model_run_{run_id}.pth")
        torch.save(similar_agent.model.state_dict(), similar_model_path)
        torch.save(isolated_agent.model.state_dict(), isolated_model_path)
        print(f"[Run {run_id}] 模型已保存(优化版 - 仅参数)")
        similar_buffer.reset_sampled_indices()
        isolated_buffer.reset_sampled_indices()
        run_data = {
            'run_id': run_id,
            'training_time': training_time,
            'total_reward': total_cumulative_reward,
            'path_rewards': path_rewards,
            'path_similarities': {},
            'path_samples': {}
        }
        all_similarities = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            path_id = path_idx + 1
            if path_id in similar_group_display:
                buffer = similar_buffer
            elif path_id in isolated_group_display:
                buffer = isolated_buffer
            else:
                continue
            high_reward_samples = buffer.get_high_reward_samples(target_path, num_samples=20)
            if high_reward_samples:
                similarities = [sim for _, _, sim, _ in high_reward_samples]
                run_data['path_similarities'][path_id] = {
                    'avg_similarity': np.mean(similarities),
                    'max_similarity': np.max(similarities),
                    'min_similarity': np.min(similarities),
                    'sample_count': len(similarities)
                }
                run_data['path_samples'][path_id] = high_reward_samples
                all_similarities.extend(similarities)
            else:
                run_data['path_similarities'][path_id] = {
                    'avg_similarity': 0.0,
                    'max_similarity': 0.0,
                    'min_similarity': 0.0,
                    'sample_count': 0
                }
                run_data['path_samples'][path_id] = []
        if all_similarities:
            run_data['overall_avg_similarity'] = np.mean(all_similarities)
            run_data['max_similarity'] = np.max(all_similarities)
            run_data['min_similarity'] = np.min(all_similarities)
        else:
            run_data['overall_avg_similarity'] = 0.0
            run_data['max_similarity'] = 0.0
            run_data['min_similarity'] = 0.0
        all_runs_data.append(run_data)
        print(f"[Run {run_id}] 完成！总体平均相似度: {run_data['overall_avg_similarity']:.4f}")
        print(f"{'=' * 60}\n")
    total_time = time.time() - total_start_time
    print("\n生成综合Excel报告...")
    create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir)
    print("\n" + "=" * 60)
    print("20次运行全部完成！ - 3分钟版本 (威胁分析)")
    print("=" * 60)
    print(f"训练规模:")
    print(f"   每路径: 5轮 x 4批 x 50样本 x 3步 = 3000步/路径")
    print(f"   样本生成: 2000候选 -> 200最终样本")
    print(f"   模型保存: 仅保存模型参数(优化版)")
    print(f"   总耗时: {total_time:.2f} 秒 ({total_time / 60:.2f} 分钟)")
    print(f"   平均每次运行耗时: {total_time / 20:.2f} 秒")
    print(f"\n平均相似度统计:")
    avg_similarities = [r['overall_avg_similarity'] for r in all_runs_data]
    print(f"   总体平均: {np.mean(avg_similarities):.4f}")
    print(f"   最大值: {np.max(avg_similarities):.4f}")
    print(f"   最小值: {np.min(avg_similarities):.4f}")
    print(f"   标准差: {np.std(avg_similarities):.4f}")
    print(f"\n所有结果已保存至: {output_dir}")
    print("=" * 60)

if __name__ == "__main__":
    run_20_times_training()